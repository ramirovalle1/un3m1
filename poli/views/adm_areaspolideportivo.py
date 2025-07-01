# -*- coding: UTF-8 -*-
import json
import random
import sys

from django.contrib.contenttypes.models import ContentType
from openpyxl import workbook as openxl
from openpyxl.chart import ScatterChart, Reference, Series, PieChart, BarChart
from openpyxl.styles import Font as openxlFont
from openpyxl.styles.alignment import Alignment as alin

from xlwt import *

import openpyxl
import requests
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Q
import xlwt
from django.forms import model_to_dict
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.template import Context
from django.template.loader import get_template
from django.shortcuts import render, redirect

from balcon.models import EncuestaProceso, PreguntaEncuestaProceso
from decorators import secure_module
from sagest.funciones import encrypt_id, crear_editar_encuesta
from sagest.models import DistributivoPersona
from settings import EMAIL_DOMAIN
from sga.commonviews import adduserdata
from sga.templatetags.sga_extras import encrypt
from poli.forms import AreaPolideportivoForm, FotoAreaPolideportivoForm, ActividadPolideportivoForm, \
    SeccionPolideportivoForm, HorarioActividadPolideportivoForm, TurnoPolideportivoForm, ImplementosActividadForm, \
    DisciplinaDeportivaForm, ImplementoForm, InstructorPolideportivoForm, InstructorActividadPolideportivoForm, \
    SancionPolideportivoForm, PoliticaPolideportivoForm, PerfilesActividadForm, ClubPoliForm, IntegranteClubForm, DescuentoValorActividadForm, EncuestaPreguntaForm, TituloWebSiteForm, \
    CuerpoWebSiteForm, NoticiaDeportivaForm, PlanificacionActividadForm
from sga.funciones import MiPaginador, log, generar_nombre, remover_caracteres_especiales_unicode
from sga.models import Administrativo, Persona, DisciplinaDeportiva, Utensilios
from poli.models import *
from django.db.models import Sum, Q, F, FloatField

from utils.filtros_genericos import filtro_persona_select


@login_required(redirect_field_name='ret', login_url='/loginsga')
@secure_module
@transaction.atomic()
def view(request):
    data = {}
    adduserdata(request, data)
    usuario = request.user
    persona = request.session['persona']
    periodo = request.session['periodo']
    perfilprincipal = request.session['perfilprincipal']
    if request.method == 'POST':
        res_json = []
        action = request.POST['action']

        if action == 'mostrarsancion':
            try:
                with transaction.atomic():
                    sancion = SancionPolideportivo.objects.get(id=request.POST['id'])
                    sancion.mostrar = eval(request.POST['val'])
                    sancion.save(request)
                    log(u'Edito Sancion Polideportivo: %s' % sancion, request, "edit")
                    return JsonResponse({"result": False}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        if action == 'mostrarpolitica':
            try:
                with transaction.atomic():
                    politica = PoliticaPolideportivo.objects.get(id=request.POST['id'])
                    politica.mostrar = eval(request.POST['val'])
                    politica.save(request)
                    log(u'Edito Politica Polideportivo: %s' % politica, request, "edit")
                    return JsonResponse({"result": False}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        if action == 'addarea':
            try:
                with transaction.atomic():
                    if AreaPolideportivo.objects.filter(nombre=request.POST['nombre'], status=True).exists():
                        transaction.set_rollback(True)
                        return JsonResponse({'error': True, "message": 'Área ya existe.'}, safe=False)
                    form = AreaPolideportivoForm(request.POST)
                    if form.is_valid():
                        instance = AreaPolideportivo(nombre=form.cleaned_data['nombre'],
                                                     en_mantenimiento=form.cleaned_data['en_mantenimiento'],
                                                     descripcion=form.cleaned_data['descripcion'],
                                                     numdias=form.cleaned_data['numdias'])
                        instance.save(request)
                        if 'portada' in request.FILES:
                            newfile = request.FILES['portada']
                            newfile._name = generar_nombre(instance.nombre_input(), newfile._name)
                            instance.portada = newfile
                        if 'fondo' in request.FILES:
                            newfile2 = request.FILES['fondo']
                            newfile2._name = generar_nombre('fondo_{}'.format(instance.nombre_input()), newfile2._name)
                            instance.fondo = newfile2
                        instance.save(request)
                        log(u'Adiciono Área Polideportivo: %s' % instance, request, "add")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()], "message": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'addturno':
            try:
                with transaction.atomic():
                    instance = None
                    if TurnoPolideportivo.objects.filter(status=True, comienza=request.POST['comienza'], termina=request.POST['termina']).exists():
                        transaction.set_rollback(True)
                        return JsonResponse({'result': True, "mensaje": 'El turno que intenta ingresar ya existe.'}, safe=False)
                    else:
                        form = TurnoPolideportivoForm(request.POST)
                        if form.is_valid():
                            instance = TurnoPolideportivo(turno=form.cleaned_data['turno'],
                                                          comienza=form.cleaned_data['comienza'],
                                                          termina=form.cleaned_data['termina'],
                                                          mostrar=form.cleaned_data['mostrar'])

                            instance.save(request)
                        else:
                            transaction.set_rollback(True)
                            return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                                 "mensaje": "Error en el formulario"})
                    log(u'Adiciono Turno Polideportivo: %s' % instance, request, "add")
                    return JsonResponse({"result": False}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, "mensaje": 'Intentelo más tarde.'}, safe=False)

        elif action == 'addinstructor':
            try:
                with transaction.atomic():
                    form = InstructorPolideportivoForm(request.POST)
                    if form.is_valid():
                        if InstructorPolideportivo.objects.filter(persona=form.cleaned_data['persona']).exists():
                            return JsonResponse({'result': True, "mensaje": 'Registro ya existe.'}, safe=False)

                        instance = InstructorPolideportivo(persona=form.cleaned_data['persona'],
                                                           activo=form.cleaned_data['activo'],
                                                           descripcion=form.cleaned_data['descripcion']
                                                           )
                        instance.save(request)
                        log(u'Adiciono Turno Polideportivo: %s' % instance, request, "add")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        return JsonResponse({'result': 'bad', "form": [{k: v[0]} for k, v in form.errors.items()], "mensaje": "Error en el formulario"})

            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, "mensaje": 'Intentelo más tarde.'}, safe=False)

        elif action == 'editinstructor':
            try:
                with transaction.atomic():
                    id = encrypt_id(request.POST['id'])
                    instance=InstructorPolideportivo.objects.get(id=id)
                    form = InstructorPolideportivoForm(request.POST)
                    if not form.is_valid():
                        return JsonResponse({'result': 'bad', "form": [{k: v[0]} for k, v in form.errors.items()], "mensaje": "Error en el formulario"})

                    if InstructorPolideportivo.objects.filter(persona=form.cleaned_data['persona']).exclude(id=instance.id).exists():
                        return JsonResponse({'result': True, "mensaje": 'Registro ya existe.'}, safe=False)

                    instance.persona=form.cleaned_data['persona']
                    instance.activo=form.cleaned_data['activo']
                    instance.descripcion=form.cleaned_data['descripcion']
                    instance.save(request)
                    log(u'Edito instructor : %s' % instance, request, "edit")
                    return JsonResponse({"result": False}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, "mensaje": 'Intentelo más tarde.'}, safe=False)

        elif action == 'addinstructoract':
            try:
                with transaction.atomic():
                    actividad_id = int(request.POST['idpadre'])
                    form = InstructorActividadPolideportivoForm(request.POST)
                    if form.is_valid():
                        if InstructorActividadPolideportivo.objects.filter(instructor_id=form.cleaned_data['instructor'], actividad_id=actividad_id, status=True).exists():
                            return JsonResponse({'result': True, "message": 'Registro ya existe.'}, safe=False)

                        instance = InstructorActividadPolideportivo(instructor_id=form.cleaned_data['instructor'],
                                                                    activo=form.cleaned_data['activo'],
                                                                    actividad_id=actividad_id
                                                                    )
                        instance.save(request)
                        log(u'Adiciono Turno Polideportivo: %s' % instance, request, "add")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()], "message": "Error en el formulario"})

            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, "message": 'Intentelo más tarde.'}, safe=False)

        elif action == 'adddis':
            try:
                with transaction.atomic():
                    if DisciplinaDeportiva.objects.filter(descripcion=request.POST['descripcion'], status=True).exists():
                        transaction.set_rollback(True)
                        return JsonResponse({'error': True, "message": 'Disciplina ya existe.'}, safe=False)
                    form = DisciplinaDeportivaForm(request.POST)
                    if form.is_valid():
                        instance = DisciplinaDeportiva(descripcion=request.POST['descripcion'])
                        instance.save(request)
                        log(u'Adiciono disciplina Polideportivo: %s' % instance, request, "add")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()], "message": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'addactividad':
            try:
                with transaction.atomic():
                    if ActividadPolideportivo.objects.filter(nombre=request.POST['nombre'], area_id=request.POST['idpadre'], status=True).exists():
                        transaction.set_rollback(True)
                        return JsonResponse({'error': True, "message": 'Actividad ya existe.'}, safe=False)
                    form = ActividadPolideportivoForm(request.POST)
                    form.validar_tipo()
                    if form.is_valid():
                        instance = ActividadPolideportivo(area_id=request.POST['idpadre'],
                                                          disciplina=form.cleaned_data['disciplina'],
                                                          nombre=form.cleaned_data['nombre'],
                                                          mostrar=form.cleaned_data['mostrar'],
                                                          interno=form.cleaned_data['interno'],
                                                          externo=form.cleaned_data['externo'],
                                                          cupo=form.cleaned_data['cupo'],
                                                          valor=form.cleaned_data['valor'],
                                                          responsable=form.cleaned_data['responsable'],
                                                          fechainicio=form.cleaned_data['fechainicio'],
                                                          fechafin=form.cleaned_data['fechafin'],
                                                          descripcion=form.cleaned_data['descripcion'],
                                                          tipoactividad=form.cleaned_data['tipoactividad'],
                                                          numacompanantes=form.cleaned_data['numacompanantes'])
                        instance.save(request)
                        for carrera in form.cleaned_data['carrera']:
                            instance.carreras.add(carrera)

                        if 'portada' in request.FILES:
                            newfile2 = request.FILES['portada']
                            newfile2._name = generar_nombre('portada_{}'.format(instance.nombre_input()), newfile2._name)
                            instance.portada = newfile2
                        instance.save(request)
                        log(u'Adiciono Área Polideportivo: %s' % instance, request, "add")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()], "mensaje": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'addsancion':
            try:
                with transaction.atomic():
                    form = SancionPolideportivoForm(request.POST)
                    if form.is_valid():
                        instance = SancionPolideportivo(
                            nombre=form.cleaned_data['nombre'],
                            descripcion=form.cleaned_data['descripcion'],
                            valor=form.cleaned_data['valor'],
                            mostrar=form.cleaned_data['mostrar'])
                        instance.save(request)

                        instance.save(request)
                        log(u'Adicionó sanción Polideportivo: %s' % instance, request, "add")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()], "message": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'addpolitica':
            try:
                with transaction.atomic():
                    form = PoliticaPolideportivoForm(request.POST)
                    if form.is_valid():
                        instance = PoliticaPolideportivo(
                            nombre=form.cleaned_data['nombre'],
                            descripcion=form.cleaned_data['descripcion'],
                            mostrar=form.cleaned_data['mostrar'],
                            general=form.cleaned_data['general'])
                        instance.save(request)
                        for area in form.cleaned_data['area']:
                            instance.area.add(area)
                        log(u'Adicionó política Polideportivo: %s' % instance, request, "add")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()], "message": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'addimplementoact':
            try:
                with transaction.atomic():
                    actividad = ActividadPolideportivo.objects.get(pk=request.POST['idpadre'])
                    if ImplementosActividad.objects.filter(actividad=actividad,
                                                           utensilio=request.POST['utensilio'],
                                                           status=True).exists():
                        transaction.set_rollback(True)
                        return JsonResponse({'error': True, "message": 'Actividad ya existe.'}, safe=False)
                    form = ImplementosActividadForm(request.POST)
                    if form.is_valid():
                        instance = ImplementosActividad(actividad=actividad,
                                                        utensilio=form.cleaned_data['utensilio'],
                                                        activo=form.cleaned_data['activo'],
                                                        cantidad=form.cleaned_data['cantidad'])
                        instance.save(request)
                        log(u'Adicionó implementos en actividad Polideportivo: %s' % instance, request, "add")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()], "message": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'addperfil':
            try:
                with transaction.atomic():
                    actividad = ActividadPolideportivo.objects.get(pk=request.POST['idpadre'])
                    if actividad.perfilesactividad_set.filter(perfil=request.POST['perfil'], status=True).exists():
                        raise NameError('Perfil ya existe.')
                    form = PerfilesActividadForm(request.POST)
                    if form.is_valid():
                        instance = PerfilesActividad(actividad=actividad,
                                                     activo=form.cleaned_data['activo'],
                                                     perfil=form.cleaned_data['perfil'])

                        instance.save(request)
                        for familiar in form.cleaned_data['familiares']:
                            instance.familiares.add(familiar)
                        log(u'Adicionó perfil en actividad Polideportivo: %s' % instance, request, "add")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()], "message": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": str(ex)})

        elif action == 'addimplemento':
            try:
                with transaction.atomic():

                    if Utensilios.objects.filter(descripcion=request.POST['descripcion'],
                                                 status=True).exists():
                        transaction.set_rollback(True)
                        return JsonResponse({'error': True, "message": 'Actividad ya existe.'}, safe=False)
                    form = ImplementoForm(request.POST)
                    if form.is_valid():

                        instance = Utensilios(descripcion=form.cleaned_data['descripcion'],
                                              cantidad=form.cleaned_data['cantidad'])
                        instance.save(request)
                        log(u'Adicionó implemento para polideportivo: %s' % instance, request, "add")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()], "message": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'addseccion':
            try:
                with transaction.atomic():
                    area = AreaPolideportivo.objects.get(pk=request.POST['idpadre'])
                    if SeccionPolideportivo.objects.filter(area=area, nombre=request.POST['nombre'], status=True).exists():
                        transaction.set_rollback(True)
                        return JsonResponse({'error': True, "message": 'Actividad ya existe.'}, safe=False)
                    form = SeccionPolideportivoForm(request.POST)
                    if form.is_valid():
                        instance = SeccionPolideportivo(area=area,
                                                        nombre=form.cleaned_data['nombre'],
                                                        cupo=form.cleaned_data['cupo'],
                                                        mostrar=form.cleaned_data['mostrar'])
                        if 'fondo' in request.FILES:
                            newfile2 = request.FILES['fondo']
                            newfile2._name = generar_nombre('fondo_{}'.format(instance.nombre_input()), newfile2._name)
                            instance.fondo = newfile2

                        if 'icono' in request.FILES:
                            newfile2 = request.FILES['icono']
                            newfile2._name = generar_nombre('icono_{}'.format(instance.nombre_input()), newfile2._name)
                            instance.icono = newfile2

                        instance.save(request)
                        log(u'Adicionó sección Polideportivo: %s' % instance, request, "add")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()], "message": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'editarea':
            try:
                with transaction.atomic():
                    filtro = AreaPolideportivo.objects.get(pk=request.POST['id'])
                    f = AreaPolideportivoForm(request.POST)
                    if f.is_valid():
                        filtro.nombre = f.cleaned_data['nombre']
                        if 'descripcion' in request.POST:
                            filtro.descripcion = request.POST['descripcion']
                        filtro.numdias = f.cleaned_data['numdias']
                        filtro.en_mantenimiento = f.cleaned_data['en_mantenimiento']
                        if 'portada' in request.FILES:
                            newfile = request.FILES['portada']
                            newfile._name = generar_nombre(filtro.nombre_input(), newfile._name)
                            filtro.portada = newfile
                        if 'fondo' in request.FILES:
                            newfile2 = request.FILES['fondo']
                            newfile2._name = generar_nombre('fondo_{}'.format(filtro.nombre_input()), newfile2._name)
                            filtro.fondo = newfile2
                        filtro.save(request)
                        log(u'Edito Área Polideportivo: %s' % filtro, request, "edit")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()], "message": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'editturno':
            try:
                with transaction.atomic():
                    filtro = TurnoPolideportivo.objects.get(pk=request.POST['id'])
                    f = TurnoPolideportivoForm(request.POST)
                    if f.is_valid():
                        filtro.turno = f.cleaned_data['turno']
                        filtro.comienza = f.cleaned_data['comienza']
                        filtro.termina = f.cleaned_data['termina']
                        filtro.mostrar = f.cleaned_data['mostrar']
                        filtro.save(request)
                        log(u'Edito turno Polideportivo: %s' % filtro, request, "edit")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()], "message": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'editsancion':
            try:
                with transaction.atomic():
                    filtro = SancionPolideportivo.objects.get(pk=request.POST['id'])
                    f = SancionPolideportivoForm(request.POST)
                    if f.is_valid():
                        filtro.nombre = f.cleaned_data['nombre']
                        filtro.descripcion = f.cleaned_data['descripcion']
                        filtro.valor = f.cleaned_data['valor']
                        filtro.mostrar = f.cleaned_data['mostrar']
                        filtro.save(request)

                        log(u'Edito sanción Polideportivo: %s' % filtro, request, "edit")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()], "message": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'editpolitica':
            try:
                with transaction.atomic():
                    filtro = PoliticaPolideportivo.objects.get(pk=request.POST['id'])
                    f = PoliticaPolideportivoForm(request.POST)
                    if f.is_valid():
                        filtro.nombre = request.POST['nombre']
                        filtro.descripcion = request.POST['descripcion']
                        if 'mostrar' in request.POST:
                            filtro.mostrar = True
                        else:
                            filtro.mostrar = False
                        if 'general' in request.POST:
                            filtro.general = True
                        else:
                            filtro.general = False
                        filtro.area.clear()
                        for area in f.cleaned_data['area']:
                            filtro.area.add(area)
                        filtro.save(request)
                        log(u'Edito política de Polideportivo: %s' % filtro, request, "edit")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()], "message": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'editdis':
            try:
                with transaction.atomic():
                    filtro = DisciplinaDeportiva.objects.get(pk=request.POST['id'])
                    f = DisciplinaDeportivaForm(request.POST)
                    if f.is_valid():
                        filtro.descripcion = request.POST['descripcion']
                        filtro.save(request)
                        log(u'Edito disciplina Polideportivo: %s' % filtro, request, "edit")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()], "message": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'editactividad':
            try:
                with transaction.atomic():
                    filtro = ActividadPolideportivo.objects.get(pk=request.POST['id'])
                    f = ActividadPolideportivoForm(request.POST)
                    f.validar_tipo()
                    if f.is_valid():
                        filtro.disciplina = f.cleaned_data['disciplina']
                        filtro.nombre = f.cleaned_data['nombre']
                        filtro.descripcion = f.cleaned_data['descripcion']
                        filtro.cupo = f.cleaned_data['cupo']
                        filtro.valor = f.cleaned_data['valor']
                        filtro.mostrar = f.cleaned_data['mostrar']
                        filtro.interno = f.cleaned_data['interno']
                        filtro.externo = f.cleaned_data['externo']
                        filtro.responsable_id = f.cleaned_data['responsable']
                        filtro.fechainicio = f.cleaned_data['fechainicio']
                        filtro.fechafin = f.cleaned_data['fechafin']
                        filtro.numacompanantes = f.cleaned_data['numacompanantes']
                        filtro.tipoactividad = f.cleaned_data['tipoactividad']

                        if 'portada' in request.FILES:
                            newfile2 = request.FILES['portada']
                            newfile2._name = generar_nombre('portada_{}'.format(filtro.nombre_input()), newfile2._name)
                            filtro.portada = newfile2

                        filtro.carreras.clear()
                        for carrera in f.cleaned_data['carrera']:
                            filtro.carreras.add(carrera)
                        filtro.save(request)
                        log(u'Edito actividad de Polideportivo: %s' % filtro, request, "edit")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()], "message": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'editimplementoact':
            try:
                with transaction.atomic():
                    filtro = ImplementosActividad.objects.get(pk=request.POST['id'])
                    f = ImplementosActividadForm(request.POST)
                    if f.is_valid():
                        filtro.utensilio = f.cleaned_data['utensilio']
                        filtro.cantidad = f.cleaned_data['cantidad']
                        filtro.activo = f.cleaned_data['activo']
                        filtro.save(request)
                        log(u'Editó implemento de actividad Polideportivo: %s' % filtro, request, "edit")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()], "message": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'editperfil':
            try:
                with transaction.atomic():
                    filtro = PerfilesActividad.objects.get(pk=request.POST['id'])
                    f = PerfilesActividadForm(request.POST)
                    if f.is_valid():
                        filtro.perfil = f.cleaned_data['perfil']
                        filtro.activo = f.cleaned_data['activo']
                        filtro.save(request)
                        filtro.familiares.clear()
                        for familiar in f.cleaned_data['familiares']:
                            filtro.familiares.add(familiar)
                        log(u'Editó perfil de actividad Polideportivo: %s' % filtro, request, "edit")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()], "message": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'editinstructoract':
            try:
                with transaction.atomic():
                    filtro = InstructorActividadPolideportivo.objects.get(pk=request.POST['id'])
                    f = InstructorActividadPolideportivoForm(request.POST)
                    if f.is_valid():
                        filtro.instructor_id = f.cleaned_data['instructor']
                        filtro.activo = f.cleaned_data['activo']
                        filtro.save(request)
                        log(u'Editó instructor de actividad Polideportivo: %s' % filtro, request, "edit")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()], "message": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'editimplemento':
            try:
                with transaction.atomic():
                    filtro = Utensilios.objects.get(pk=request.POST['id'])
                    f = ImplementoForm(request.POST)
                    if f.is_valid():
                        filtro.descripcion = f.cleaned_data['descripcion']
                        filtro.cantidad = f.cleaned_data['cantidad']
                        filtro.save(request)
                        log(u'Editó implemento de Polideportivo: %s' % filtro, request, "edit")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()], "message": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'editseccion':
            try:
                with transaction.atomic():
                    filtro = SeccionPolideportivo.objects.get(pk=request.POST['id'])
                    f = SeccionPolideportivoForm(request.POST)
                    if f.is_valid():
                        filtro.nombre = f.cleaned_data['nombre']
                        filtro.cupo = f.cleaned_data['cupo']
                        if 'mostrar' in request.POST:
                            filtro.mostrar = True
                        else:
                            filtro.mostrar = False
                        if 'fondo' in request.FILES:
                            newfile2 = request.FILES['fondo']
                            newfile2._name = generar_nombre('fondo_{}'.format(filtro.nombre_input()),
                                                            newfile2._name)
                            filtro.fondo = newfile2

                        if 'icono' in request.FILES:
                            newfile2 = request.FILES['icono']
                            newfile2._name = generar_nombre('icono_{}'.format(filtro.nombre_input()),
                                                            newfile2._name)
                            filtro.icono = newfile2

                        filtro.save(request)
                        log(u'Editó sección de Polideportivo: %s' % filtro, request, "edit")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()], "message": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'deletearea':
            try:
                with transaction.atomic():
                    instancia = AreaPolideportivo.objects.get(pk=int(request.POST['id']))
                    instancia.status = False
                    instancia.save(request)
                    log(u'Elimino Área Polideportivo: %s' % instancia, request, "delete")
                    res_json = {"error": False}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        elif action == 'deleteactividad':
            try:
                with transaction.atomic():
                    instancia = ActividadPolideportivo.objects.get(pk=int(request.POST['id']))
                    instancia.status = False
                    instancia.save(request)
                    log(u'Elimino Actividad Polideportivo: %s' % instancia, request, "delete")
                    res_json = {"error": False}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        elif action == 'deleteturno':
            try:
                with transaction.atomic():
                    instancia = TurnoPolideportivo.objects.get(pk=int(request.POST['id']))
                    instancia.status = False
                    instancia.save(request)
                    log(u'Elimino Turno Polideportivo: %s' % instancia, request, "delete")
                    res_json = {"error": False}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        elif action == 'deletedis':
            try:
                with transaction.atomic():
                    instancia = DisciplinaDeportiva.objects.get(pk=int(request.POST['id']))
                    instancia.status = False
                    instancia.save(request)
                    log(u'Elimino disciplina Polideportivo: %s' % instancia, request, "delete")
                    res_json = {"error": False}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        elif action == 'deleteimplemento':
            try:
                with transaction.atomic():
                    instancia = Utensilios.objects.get(pk=int(request.POST['id']))
                    instancia.status = False
                    instancia.save(request)
                    log(u'Elimino implemento de Polideportivo: %s' % instancia, request, "delete")
                    res_json = {"error": False}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        elif action == 'deleteimplementoact':
            try:
                with transaction.atomic():
                    instancia = ImplementosActividad.objects.get(pk=int(request.POST['id']))
                    instancia.status = False
                    instancia.save(request)
                    log(u'Elimino implemento de actividad: %s' % instancia, request, "delete")
                    res_json = {"error": False}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        elif action == 'delinstructor':
            try:
                with transaction.atomic():
                    instancia = InstructorPolideportivo.objects.get(pk=int(request.POST['id']))
                    instancia.status = False
                    instancia.save(request)
                    log(u'Elimino instructor: %s' % instancia, request, "delete")
                    res_json = {"error": False}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        elif action == 'deleteinstructoract':
            try:
                with transaction.atomic():
                    instancia = InstructorActividadPolideportivo.objects.get(pk=int(request.POST['id']))
                    instancia.status = False
                    instancia.save(request)
                    log(u'Eliminó instructor de actividad Polideportivo: %s' % instancia, request, "delete")
                    res_json = {"error": False}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        elif action == 'deletehorario':
            try:
                with transaction.atomic():
                    instancia = HorarioActividadPolideportivo.objects.get(pk=int(request.POST['id']))
                    instancia.status = False
                    instancia.save(request)
                    log(u'Eliminó horario de actividad Polideportivo: %s' % instancia, request, "delete")
                    res_json = {"error": False}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        elif action == 'deletesancion':
            try:
                with transaction.atomic():
                    instancia = SancionPolideportivo.objects.get(pk=int(request.POST['id']))
                    instancia.status = False
                    instancia.save(request)
                    log(u'Eliminó sanción de Polideportivo: %s' % instancia, request, "delete")
                    res_json = {"error": False}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        elif action == 'deleteseccion':
            try:
                with transaction.atomic():
                    instancia = SeccionPolideportivo.objects.get(pk=int(request.POST['id']))
                    instancia.status = False
                    instancia.save(request)
                    log(u'Eliminó sección de Polideportivo: %s' % instancia, request, "delete")
                    res_json = {"error": False}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        elif action == 'deletepolitica':
            try:
                with transaction.atomic():
                    instancia = PoliticaPolideportivo.objects.get(pk=int(request.POST['id']))
                    instancia.status = False
                    instancia.save(request)
                    log(u'Eliminó política de Polideportivo: %s' % instancia, request, "delete")
                    res_json = {"error": False}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        elif action == 'deleteperfil':
            try:
                with transaction.atomic():
                    instancia = PerfilesActividad.objects.get(pk=int(request.POST['id']))
                    instancia.status = False
                    instancia.save(request)
                    log(u'Eliminó perfil de Polideportivo: %s' % instancia, request, "delete")
                    res_json = {"error": False}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        elif action == 'enmantenimiento':
            try:
                registro = AreaPolideportivo.objects.get(pk=request.POST['id'])
                registro.en_mantenimiento = True if request.POST['val'] == 'y' else False
                registro.save(request)
                log(u'Area en Mantenimiento : %s (%s)' % (registro, registro.en_mantenimiento), request, "edit")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad"})

        elif action == 'activainstructor':
            try:
                registro = InstructorPolideportivo.objects.get(pk=request.POST['id'])
                registro.activo = True if request.POST['val'] == 'y' else False
                registro.save(request)
                log(u'Instructor activo : %s (%s)' % (registro, registro.activo), request, "edit")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad"})

        elif action == 'activaturno':
            try:
                registro = TurnoPolideportivo.objects.get(pk=request.POST['id'])
                registro.mostrar = True if request.POST['val'] == 'y' else False
                registro.save(request)
                log(u'Turno activo : %s (%s)' % (registro, registro.mostrar), request, "edit")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad"})

        elif action == 'activaperfil':
            try:
                registro = PerfilesActividad.objects.get(pk=request.POST['id'])
                registro.activo = eval(request.POST['val'].title())
                registro.save(request)
                log(u'Edito perfil : %s (%s)' % (registro, registro.activo), request, "edit")
                return JsonResponse({"result": True, 'mensaje': 'Cambios guardados'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", 'mensaje': str(ex)})

        elif action == 'activoinstructoract':
            try:
                registro = InstructorActividadPolideportivo.objects.get(pk=request.POST['id'])
                registro.activo = True if request.POST['val'] == 'y' else False
                registro.save(request)
                log(u'Instructor actividad activo : %s (%s)' % (registro, registro.activo), request, "edit")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad"})

        elif action == 'estadoimplementoact':
            try:
                registro = ImplementosActividad.objects.get(pk=request.POST['id'])
                registro.activo = True if request.POST['val'] == 'y' else False
                registro.save(request)
                log(u'cambia estado implemento actividad activo : %s (%s)' % (registro, registro.activo), request, "edit")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad"})

        elif action == 'addfoto':
            try:
                with transaction.atomic():
                    form = FotoAreaPolideportivoForm(request.POST)
                    if form.is_valid():
                        area = AreaPolideportivo.objects.get(pk=request.POST['id'])
                        instance = FotosAreaPolideportivo(area=area, orden=form.cleaned_data['orden'], visible=form.cleaned_data['visible'])
                        instance.save(request)
                        if 'foto' in request.FILES:
                            newfile = request.FILES['foto']
                            newfile._name = generar_nombre('foto_{}_{}'.format(instance.area.nombre, instance.orden), newfile._name)
                            instance.foto = newfile
                        instance.save(request)
                        log(u'Adiciono Foto Área Polideportivo: %s' % instance, request, "add")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()], "message": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'editfoto':
            try:
                with transaction.atomic():
                    filtro = FotosAreaPolideportivo.objects.get(pk=request.POST['id'])
                    f = FotoAreaPolideportivoForm(request.POST)
                    if f.is_valid():
                        filtro.orden = f.cleaned_data['orden']
                        filtro.visible = f.cleaned_data['visible']
                        filtro.save(request)
                        if 'foto' in request.FILES:
                            newfile = request.FILES['foto']
                            newfile._name = generar_nombre('foto_{}_{}'.format(filtro.area.nombre, filtro.orden), newfile._name)
                            filtro.foto = newfile
                        filtro.save(request)
                        log(u'Edito Foto Área Polideportivo: %s' % filtro, request, "edit")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()], "message": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'deletefotoarea':
            try:
                with transaction.atomic():
                    instancia = FotosAreaPolideportivo.objects.get(pk=int(request.POST['id']))
                    instancia.status = False
                    instancia.save(request)
                    log(u'Elimino Foto Área Polideportivo: %s' % instancia, request, "delete")
                    res_json = {"error": False}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        elif action == 'fotovisible':
            try:
                registro = FotosAreaPolideportivo.objects.get(pk=request.POST['id'])
                registro.visible = True if request.POST['val'] == 'y' else False
                registro.save(request)
                log(u'Foto de Area, Visible : %s (%s)' % (registro, registro.visible), request, "edit")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad"})

        elif action == 'addhorario':
            try:
                form = HorarioActividadPolideportivoForm(request.POST)
                idp = request.POST.get('planificacion', '')
                if form.is_valid():
                    filtro = Q(actividad_id=int(request.POST['actividad']))
                    if idp:
                        idp = encrypt_id(idp)
                        filtro = filtro & Q(planificacion_id=idp)
                    if HorarioActividadPolideportivo.objects.filter(filtro,
                                                                    dia=form.cleaned_data['dia'], turno=form.cleaned_data['turno'],
                                                                    fechainicio=form.cleaned_data['fechainicio'],
                                                                    fechafin=form.cleaned_data['fechafin'], status=True).exists():
                        raise NameError("Existe un horario con la misma fecha y turno.")
                    if not form.cleaned_data['fechainicio'] <= form.cleaned_data['fechafin']:
                        raise NameError("No puede ser mayor la fecha de inicio.")
                    if form.cleaned_data['fechainicio'] == form.cleaned_data['fechafin']:
                        if not int(form.cleaned_data['dia']) == form.cleaned_data['fechainicio'].weekday() + 1:
                            raise NameError("La fecha no concuerdan con el dia.")
                        if not form.cleaned_data['instructor'].actividad_id == int(request.POST['actividad']):
                            raise NameError("Instructor no pertenece a la actividad.")

                    clase = HorarioActividadPolideportivo(actividad_id=int(request.POST['actividad']),
                                                          planificacion_id=idp if idp else None,
                                                          turno=form.cleaned_data['turno'],
                                                          dia=form.cleaned_data['dia'],
                                                          fechainicio=form.cleaned_data['fechainicio'],
                                                          instructor=form.cleaned_data['instructor'],
                                                          mostrar=form.cleaned_data['mostrar'],
                                                          fechafin=form.cleaned_data['fechafin'])
                    clase.save(request)
                    log(u'Adiciono horario en actividad: %s [%s]' % (clase, clase.id), request, "add")
                    return JsonResponse({"result": False}, safe=False)
                else:
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                         "message": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": f"Error:{ex}"}, safe=False)

        elif action == 'edithorario':
            try:
                with transaction.atomic():
                    filtro = HorarioActividadPolideportivo.objects.get(pk=request.POST['id'])
                    f = HorarioActividadPolideportivoForm(request.POST)
                    if f.is_valid():
                        filtro.turno = f.cleaned_data['turno']
                        filtro.dia = f.cleaned_data['dia']
                        filtro.mostrar = f.cleaned_data['mostrar']
                        filtro.fechainicio = f.cleaned_data['fechainicio']
                        filtro.fechafin = f.cleaned_data['fechafin']
                        filtro.instructor = f.cleaned_data['instructor']
                        filtro.save(request)

                        log(u'Edito horario: %s' % filtro, request, "edit")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()], "message": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'mostraractividad':
            try:
                with transaction.atomic():
                    actividad = ActividadPolideportivo.objects.get(id=request.POST['id'])
                    actividad.mostrar = eval(request.POST['val'])
                    actividad.save(request)
                    log(u'Edito Actividad Polideportivo: %s' % actividad, request, "edit")
                    return JsonResponse({"result": False}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'mostrarseccion':
            try:
                with transaction.atomic():
                    seccion = SeccionPolideportivo.objects.get(id=int(encrypt(request.POST['id'])))
                    seccion.mostrar = eval(request.POST['val'])
                    seccion.save(request)
                    log(u'Edito mostrar seccion: %s' % seccion, request, "mostrarseccion")
                    return JsonResponse({"result": False, "mensaje": "Se actualizo mostrar sección"}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'addimplemento':
            try:
                with transaction.atomic():

                    if Utensilios.objects.filter(descripcion=request.POST['descripcion'],
                                                 status=True).exists():
                        transaction.set_rollback(True)
                        return JsonResponse({'error': True, "message": 'Actividad ya existe.'}, safe=False)
                    form = ImplementoForm(request.POST)
                    if form.is_valid():

                        instance = Utensilios(descripcion=form.cleaned_data['descripcion'],
                                              cantidad=form.cleaned_data['cantidad'])
                        instance.save(request)
                        log(u'Adicionó implemento para polideportivo: %s' % instance, request, "add")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()], "message": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'addclub':
            try:
                with transaction.atomic():
                    form = ClubPoliForm(request.POST)
                    if form.is_valid():
                        instance = ClubPoli(nombre=form.cleaned_data['nombre'],
                                            finicio=form.cleaned_data['finicio'],
                                            ffin=form.cleaned_data['ffin'],
                                            descripcion=form.cleaned_data['descripcion'],
                                            responsable=form.cleaned_data['responsable'],
                                            disciplina=form.cleaned_data['disciplina'])
                        instance.save(request)
                        log(u'Adicionó club para polideportivo: %s' % instance, request, "add")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()], "mensaje": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'editclub':
            try:
                with transaction.atomic():
                    id = int(encrypt(request.POST['id']))
                    club = ClubPoli.objects.get(pk=id)
                    form = ClubPoliForm(request.POST)
                    if form.is_valid():
                        club.nombre = form.cleaned_data['nombre']
                        club.finicio = form.cleaned_data['finicio']
                        club.ffin = form.cleaned_data['ffin']
                        club.descripcion = form.cleaned_data['descripcion']
                        club.responsable = form.cleaned_data['responsable']
                        club.disciplina = form.cleaned_data['disciplina']
                        club.save(request)
                        log(u'Edito club para polideportivo: %s' % club, request, "editclub")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()], "mensaje": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'delclub':
            try:
                with transaction.atomic():
                    instancia = ClubPoli.objects.get(pk=int(encrypt(request.POST['id'])))
                    instancia.status = False
                    instancia.save(request)
                    log(u'Eliminó club de Polideportivo: %s' % instancia, request, "delclub")
                    res_json = {"error": False}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        elif action == 'addintegrante':
            try:
                with transaction.atomic():
                    form = IntegranteClubForm(request.POST)
                    id = int(encrypt(request.POST['idpadre']))
                    if form.is_valid():
                        instance = IntegranteClubPoli(club_id=id,
                                                      integrante=form.cleaned_data['integrante'],
                                                      tipointegrante=form.cleaned_data['tipointegrante'])
                        instance.save(request)
                        log(u'Adicionó integrante para club: %s' % instance, request, "addintegrante")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()], "mensaje": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'delintegrante':
            try:
                with transaction.atomic():
                    instancia = IntegranteClubPoli.objects.get(pk=int(encrypt(request.POST['id'])))
                    instancia.status = False
                    instancia.save(request)
                    log(u'Eliminó integrante de club: %s' % instancia, request, "delintegrante")
                    res_json = {"error": False}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        elif action == 'adddescuento':
            with transaction.atomic():
                try:
                    idperfil = int(encrypt(request.POST['id']))
                    form = DescuentoValorActividadForm(request.POST)
                    aplica = request.POST['aplicavigencia'] if 'aplicavigencia' in request.POST else None
                    if not aplica:
                        form.fields['fechainicio'].required = False
                        form.fields['fechafin'].required = False
                    if form.is_valid():
                        descuento = DescuentoValorActividad(perfilactividad_id=idperfil,
                                                            porcentaje=form.cleaned_data['porcentaje'],
                                                            valor_descuento=form.cleaned_data['valor_descuento'],
                                                            valor_final=form.cleaned_data['valor_final'],
                                                            fechainicio=form.cleaned_data['fechainicio'],
                                                            fechafin=form.cleaned_data['fechafin'],
                                                            aplicavigencia=form.cleaned_data['aplicavigencia'], )
                        descuento.save(request)
                        log(u'Agrego descuento: %s' % descuento, request, "add")
                        diccionario = {'id': descuento.id,
                                       'porcentaje': descuento.porcentaje,
                                       'valor_descuento': descuento.valor_descuento,
                                       'valor_final': descuento.valor_final,
                                       'fechainicio': descuento.fechainicio.strftime('%Y-%m-%d') if aplica else None,
                                       'fechafin': descuento.fechafin.strftime('%Y-%m-%d') if aplica else None,
                                       'publicado': descuento.publicado,
                                       'aplicavigencia': descuento.aplicavigencia,
                                       }
                        return JsonResponse({'result': True, 'data_return': True, 'mensaje': u'Guardado con éxito', 'data': diccionario, })
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                             "mensaje": "Error en el formulario"})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "mensaje": 'Error: {}'.format(ex)}, safe=False)

        elif action == 'editdescuento':
            with transaction.atomic():
                try:
                    id = int(request.POST['iddescuento'])
                    form = DescuentoValorActividadForm(request.POST)
                    aplica = request.POST['aplicavigencia'] if 'aplicavigencia' in request.POST else None
                    if not aplica:
                        form.fields['fechainicio'].required = False
                        form.fields['fechafin'].required = False
                    if form.is_valid():
                        descuento = DescuentoValorActividad.objects.get(id=id)
                        descuento.porcentaje = form.cleaned_data['porcentaje']
                        descuento.valor_descuento = form.cleaned_data['valor_descuento']
                        descuento.valor_final = form.cleaned_data['valor_final']
                        descuento.fechainicio = form.cleaned_data['fechainicio']
                        descuento.fechafin = form.cleaned_data['fechafin']
                        descuento.aplicavigencia = form.cleaned_data['aplicavigencia']
                        descuento.save(request)
                        log(u'Edito descuento: %s' % descuento, request, "edit")
                        diccionario = {'id': descuento.id,
                                       'porcentaje': descuento.porcentaje,
                                       'valor_descuento': descuento.valor_descuento,
                                       'valor_final': descuento.valor_final,
                                       'fechainicio': descuento.fechainicio.strftime('%Y-%m-%d') if aplica else None,
                                       'fechafin': descuento.fechafin.strftime('%Y-%m-%d') if aplica else None,
                                       'publicado': descuento.publicado,
                                       'aplicavigencia': descuento.aplicavigencia,
                                       'edit': True,
                                       }
                        return JsonResponse({'result': True, 'data_return': True, 'mensaje': u'Editado con éxito', 'data': diccionario})
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                             "mensaje": "Error en el formulario"})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "mensaje": 'Error: {}'.format(ex)}, safe=False)

        elif action == 'publicardescuento':
            with transaction.atomic():
                try:
                    iddescuento = int(request.POST['id'])
                    idperfil = int(request.POST['idp'])
                    estado = eval(request.POST['val'].title())
                    if estado:
                        descuentos = DescuentoValorActividad.objects.filter(status=True, perfilactividad_id=idperfil).exclude(id=iddescuento)
                        for d in descuentos:
                            d.publicado = False
                            d.save(request)
                    descuento = DescuentoValorActividad.objects.get(id=iddescuento)
                    descuento.publicado = estado
                    descuento.save(request)
                    log(u'Publico descuento: %s (%s)' % (descuento, descuento.publicado), request, "edit")
                    return JsonResponse({'result': True, 'mensaje': 'Cambios guardados'})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': False, "mensaje": 'Error: {}'.format(ex)})

        elif action == 'deldescuento':
            with transaction.atomic():
                try:
                    instancia = DescuentoValorActividad.objects.get(pk=int(request.POST['id']))
                    instancia.status = False
                    instancia.save(request)
                    log(u'Elimino descuento de actividad: %s' % instancia, request, "del")
                    res_json = {"error": False, "mensaje": 'Registro eliminado'}
                except Exception as ex:
                    res_json = {'error': True, "message": "Error: {}".format(ex)}
                return JsonResponse(res_json, safe=False)

        elif action == 'addencuesta':
            with transaction.atomic():
                try:
                    id = encrypt_id(request.POST['id'])
                    form = EncuestaPreguntaForm(request.POST)
                    if not form.is_valid():
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()], "message": "Error en el formulario"})

                    actividad = ActividadPolideportivo.objects.get(id=id)
                    encuesta = crear_editar_encuesta(request, actividad, form, 1)
                    preguntas = json.loads(request.POST['lista_items1'])
                    lista = []
                    for p in preguntas:
                        idpregunta = p['id_pregunta']
                        if not idpregunta:
                            pregunta = PreguntaEncuestaProceso(encuesta=encuesta,
                                                               estado=p['activo'],
                                                               descripcion=p['pregunta'])
                            pregunta.save(request)
                            log(u'Agrego pregunta : %s' % pregunta, request, "add")
                        else:
                            pregunta = PreguntaEncuestaProceso.objects.get(id=idpregunta)
                            if not pregunta.en_uso():
                                pregunta.descripcion = p['pregunta']
                            pregunta.estado = p['activo']
                            pregunta.save(request)
                            log(u'Edito pregunta : %s' % pregunta, request, "edit")
                        lista.append(pregunta.id)
                    PreguntaEncuestaProceso.objects.filter(status=True, encuesta=encuesta).exclude(id__in=lista).update(status=False)
                    return JsonResponse({'result': False, 'mensaje': u'Guardado con éxito'})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "mensaje": 'Error: {}'.format(ex)}, safe=False)

        # SITIO WEB UNEMI DEPORTE
        elif action == 'editsecciontitle':
            with transaction.atomic():
                try:
                    id = encrypt_id(request.POST['id'])
                    form = TituloWebSiteForm(request.POST, request.FILES)
                    if not form.is_valid():
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()], "message": "Error en el formulario"})

                    instancia = TituloWebSite.objects.get(id=id)
                    newfile = request.FILES.get('fondotitulo', None)
                    if newfile:
                        extension = newfile._name.split('.')
                        exte = extension[len(extension) - 1]
                        if newfile.size > 2194304:
                            raise NameError(u"El tamaño del archivo es mayor a 2 Mb.")
                        if not exte.lower() in ['png', 'jpg', 'jpeg']:
                            raise NameError(u"Solo se permite archivos de formato .png, .jpg, .jpeg")
                        newfile._name = generar_nombre(f'Fondo_', newfile._name)
                        instancia.fondotitulo = newfile
                    instancia.titulo = form.cleaned_data['titulo']
                    instancia.subtitulo = form.cleaned_data['subtitulo']
                    instancia.publicado = form.cleaned_data['publicado']
                    instancia.save(request)
                    log(f'Edito titulo de cabecera {instancia}', request, 'edit')
                    return JsonResponse({'result': False, 'to':f'{request.path}?action=sitioweb&s_activa={instancia.seccion}' ,'mensaje': u'Guardado con éxito'})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "mensaje": 'Error: {}'.format(ex)}, safe=False)

        elif action == 'addcuerpo':
            with transaction.atomic():
                try:
                    idp = encrypt_id(request.POST['idp'])
                    form = CuerpoWebSiteForm(request.POST)
                    if not form.is_valid():
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()], "message": "Error en el formulario"})

                    instancia = CuerpoWebSite(titulowebsite_id=idp,
                                              titulo=form.cleaned_data['titulo'],
                                              descripcion=form.cleaned_data['descripcion'],
                                              ubicacion=form.cleaned_data['ubicacion'],
                                              orden=form.cleaned_data['orden'],
                                              publicado=form.cleaned_data['publicado'],
                                              )
                    instancia.save(request)
                    log(f'Adiciono cuerpo de sitio web {instancia}', request, 'add')
                    return JsonResponse({'result': False,'to':f'{request.path}?action=sitioweb&s_activa={instancia.titulowebsite.seccion}', 'mensaje': u'Guardado con éxito'})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "mensaje": 'Error: {}'.format(ex)}, safe=False)

        elif action == 'editcuerpo':
            with transaction.atomic():
                try:
                    id = encrypt_id(request.POST['id'])
                    form = CuerpoWebSiteForm(request.POST)
                    if not form.is_valid():
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()], "message": "Error en el formulario"})
                    instancia = CuerpoWebSite.objects.get(id=id)
                    instancia.titulo = form.cleaned_data['titulo']
                    instancia.descripcion = form.cleaned_data['descripcion']
                    # instancia.ubicacion = form.cleaned_data['ubicacion']
                    instancia.orden = form.cleaned_data['orden']
                    instancia.publicado = form.cleaned_data['publicado']
                    instancia.save(request)
                    log(f'Edito cuerpo de sitio web {instancia}', request, 'edit')
                    return JsonResponse({'result': False, 'to': f'{request.path}?action=sitioweb&s_activa={instancia.titulowebsite.seccion}', 'mensaje': u'Guardado con éxito'})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "mensaje": 'Error: {}'.format(ex)}, safe=False)

        elif action == 'delcuerpo':
            try:
                with transaction.atomic():
                    instancia = CuerpoWebSite.objects.get(pk=int(encrypt(request.POST['id'])))
                    instancia.status = False
                    instancia.save(request)
                    log(u'Eliminó Cuerpo de sitio web: %s' % instancia, request, "del")
                    res_json = {"error": False,'to': f'{request.path}?action=sitioweb&s_activa={instancia.titulowebsite.seccion}'}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        #NOTICIAS

        elif action == 'addnoticia':
            with transaction.atomic():
                try:
                    form = NoticiaDeportivaForm(request.POST, request.FILES)
                    if not form.is_valid():
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()], "mensaje": "Error en el formulario, revise todos los campos"})

                    instancia = NoticiaDeportiva(titulo=form.cleaned_data['titulo'],
                                                  subtitulo=form.cleaned_data['subtitulo'],
                                                  descripcion=form.cleaned_data['descripcion'],
                                                  principal=form.cleaned_data['principal'],
                                                  portada=form.cleaned_data['portada'],
                                                  publicado=form.cleaned_data['publicado'],
                                                  )
                    instancia.save(request)
                    log(f'Adiciono noticia nueva {instancia}', request, 'add')
                    return JsonResponse({'result': False, 'mensaje': u'Guardado con éxito'})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "mensaje": 'Error: {}'.format(ex)}, safe=False)

        elif action == 'editnoticia':
            with transaction.atomic():
                try:
                    instancia = NoticiaDeportiva.objects.get(id=encrypt_id(request.POST['id']))
                    form = NoticiaDeportivaForm(request.POST, request.FILES, instancia=instancia)
                    if not form.is_valid():
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()], "message": "Error en el formulario"})

                    instancia.titulo = form.cleaned_data['titulo']
                    instancia.subtitulo = form.cleaned_data['subtitulo']
                    instancia.descripcion = form.cleaned_data['descripcion']
                    instancia.principal = form.cleaned_data['principal']
                    instancia.portada = form.cleaned_data['portada']
                    instancia.publicado = form.cleaned_data['publicado']
                    instancia.save(request)
                    log(f'Edito noticia {instancia}', request, 'edit')
                    return JsonResponse({'result': False, 'mensaje': u'Guardado con éxito'})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "mensaje": 'Error: {}'.format(ex)}, safe=False)

        elif action == 'publicarnoticia':
            with transaction.atomic():
                try:
                    instancia = NoticiaDeportiva.objects.get(id=encrypt_id(request.POST['id']))
                    if not instancia.publicado and not instancia.puede_cambiar_tipo_original():
                        raise NameError('No puede se puede publicar noticia principal por que solo se admiten hasta 3 noticias principales y ya existen 3 noticias principales vigentes')
                    instancia.publicado = not instancia.publicado
                    instancia.save(request)
                    log(f'Edito noticia {instancia}', request, 'edit')
                    return JsonResponse({'result': 'ok', 'mensaje': u'Guardado con éxito'})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "mensaje": 'Error: {}'.format(ex)}, safe=False)

        elif action == 'cambiartipo':
            with transaction.atomic():
                try:
                    instancia = NoticiaDeportiva.objects.get(id=encrypt_id(request.POST['id']))
                    if not instancia.puede_cambiar_tipo():
                        raise NameError('No puede ser una noticia principal por que solo se admiten hasta 3 noticias principales y ya existen 3 noticias principales vigentes')
                    instancia.principal = not instancia.principal
                    instancia.save(request)
                    log(f'Edito noticia {instancia}', request, 'edit')
                    return JsonResponse({'result': 'ok', 'mensaje': u'Guardado con éxito'})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "mensaje": 'Error: {}'.format(ex)}, safe=False)

        elif action == 'delnoticia':
            try:
                with transaction.atomic():
                    instancia = NoticiaDeportiva.objects.get(pk=encrypt_id(request.POST['id']))
                    instancia.status = False
                    instancia.save(request)
                    log(u'Eliminó noticia: %s' % instancia, request, "del")
                    res_json = {"error": False}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

            # NOTICIAS

        elif action == 'addplanificacion':
            with transaction.atomic():
                try:
                    actividad = ActividadPolideportivo.objects.get(id=encrypt_id(request.POST['idp']))
                    form = PlanificacionActividadForm(request.POST, request.FILES)
                    if not form.is_valid():
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()], "mensaje": "Error en el formulario, revise todos los campos"})
                    if form.cleaned_data['activo']:
                        actividad.planificacionactividad_set.filter(status=True, activo=True).update(activo=False)
                    instancia = PlanificacionActividad(actividad=actividad,
                                                 nombre=form.cleaned_data['nombre'],
                                                 fechainicio=form.cleaned_data['fechainicio'],
                                                 fechafin=form.cleaned_data['fechafin'],
                                                 costo=form.cleaned_data['costo'],
                                                 cupo=form.cleaned_data['cupo'],
                                                 activo=form.cleaned_data['activo'])
                    instancia.save(request)
                    log(f'Adiciono planificación de actividad deportiva: {instancia}', request, 'add')
                    return JsonResponse({'result': False, 'mensaje': u'Guardado con éxito'})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "mensaje": 'Error: {}'.format(ex)}, safe=False)

        elif action == 'editplanificacion':
            with transaction.atomic():
                try:
                    instancia = PlanificacionActividad.objects.get(id=encrypt_id(request.POST['id']))
                    form = PlanificacionActividadForm(request.POST, request.FILES, instancia=instancia)
                    if not form.is_valid():
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()], "message": "Error en el formulario"})

                    instancia.nombre = form.cleaned_data['nombre']
                    instancia.fechainicio = form.cleaned_data['fechainicio']
                    instancia.fechafin = form.cleaned_data['fechafin']
                    instancia.costo = form.cleaned_data['costo']
                    instancia.cupo = form.cleaned_data['cupo']
                    instancia.activo = form.cleaned_data['activo']
                    instancia.save(request)
                    log(f'Edito planificación de actividad deportiva: {instancia}', request, 'edit')
                    return JsonResponse({'result': False, 'mensaje': u'Guardado con éxito'})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "mensaje": 'Error: {}'.format(ex)}, safe=False)

        elif action == 'activarplanificacion':
            with transaction.atomic():
                try:
                    instancia = PlanificacionActividad.objects.get(id=encrypt_id(request.POST['id']))
                    if not instancia.activo:
                        instancia.desactivar_planificaciones()
                    instancia.activo = not instancia.activo
                    instancia.save(request)
                    log(f'Edito planificacio: {instancia}', request, 'edit')
                    return JsonResponse({'result': True, 'mensaje': u'Guardado con éxito'})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': False, "mensaje": 'Error: {}'.format(ex)}, safe=False)

        elif action == 'delplanificacion':
            try:
                with transaction.atomic():
                    instancia = PlanificacionActividad.objects.get(pk=encrypt_id(request.POST['id']))
                    instancia.status = False
                    instancia.save(request)
                    log(u'Eliminó planificacion deportiva: %s' % instancia, request, "del")
                    res_json = {"error": False}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        return JsonResponse({"result": "bad", "mensaje": u"Solicitud Incorrecta."})
    else:
        if 'action' in request.GET:
            data['action'] = action = request.GET['action']

            if action == 'addarea':
                try:
                    form = AreaPolideportivoForm()
                    data['form'] = form
                    template = get_template("adm_areaspolideportivo/modal/formarea.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'addactividad':
                try:
                    form = ActividadPolideportivoForm()
                    data['form'] = form
                    data['idpadre'] = pk = request.GET['idpadre']
                    template = get_template("adm_areaspolideportivo/modal/formactividad.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'addturno':
                try:
                    form = TurnoPolideportivoForm()
                    data['form'] = form
                    template = get_template("adm_areaspolideportivo/modal/formturno.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'adddis':
                try:
                    form = DisciplinaDeportivaForm()
                    data['form'] = form
                    template = get_template("adm_areaspolideportivo/modal/formdisciplina.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'addimplementoact':
                try:
                    data['idpadre'] = request.GET['idpadre']
                    form = ImplementosActividadForm()
                    data['form'] = form
                    template = get_template("adm_areaspolideportivo/modal/formimplemento.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'addperfil':
                try:
                    data['idpadre'] = request.GET['idpadre']
                    form = PerfilesActividadForm()
                    data['form'] = form
                    template = get_template("adm_areaspolideportivo/modal/formperfil.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'addimplemento':
                try:
                    form = ImplementoForm()
                    data['form'] = form
                    template = get_template("adm_areaspolideportivo/modal/formutensilio.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'addseccion':
                try:
                    data['idpadre'] = request.GET['idpadre']
                    form = SeccionPolideportivoForm()
                    data['form'] = form
                    template = get_template("adm_areaspolideportivo/modal/formseccion.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass


            elif action == 'addinstructor':
                try:
                    form = InstructorPolideportivoForm()
                    form.fields['persona'].queryset = Persona.objects.none()
                    data['form'] = form
                    template = get_template("adm_areaspolideportivo/modal/forminstructor.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'editinstructor':
                try:
                    data['filtro']= instance = InstructorPolideportivo.objects.get(id=encrypt_id(request.GET['id']))
                    form = InstructorPolideportivoForm(initial=model_to_dict(instance))
                    form.fields['persona'].queryset = Persona.objects.filter(id=instance.persona.id)
                    data['form'] = form
                    template = get_template("adm_areaspolideportivo/modal/forminstructor.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'addsancion':
                try:
                    form = SancionPolideportivoForm()
                    data['form'] = form
                    template = get_template("adm_areaspolideportivo/modal/formsancion.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'addpolitica':
                try:
                    form = PoliticaPolideportivoForm()
                    data['form'] = form
                    template = get_template("adm_areaspolideportivo/modal/formpolitica.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'addinstructoract':
                try:
                    data['idpadre'] = request.GET['idpadre']
                    form = InstructorActividadPolideportivoForm()
                    data['form'] = form
                    template = get_template("adm_areaspolideportivo/modal/forminstructoract.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'addhorario':
                try:
                    data['title'] = u'Adicionar horario'
                    actividad = ActividadPolideportivo.objects.get(pk=int(request.GET['actividad']))
                    form = HorarioActividadPolideportivoForm(initial={
                        'turno': TurnoPolideportivo.objects.get(pk=request.GET['turno']),
                        'dia': request.GET['dia']})
                    form.cargar_instructores(actividad)
                    data['actividad'] = request.GET['actividad']
                    idp=request.GET.get('id','')
                    if idp:
                        data['planificacion'] = idp
                    data['form'] = form
                    template = get_template("adm_areaspolideportivo/modal/formhorario.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'editarea':
                try:
                    data['id'] = request.GET['id']
                    data['filtro'] = filtro = AreaPolideportivo.objects.get(pk=request.GET['id'])
                    form = AreaPolideportivoForm(initial=model_to_dict(filtro))
                    data['form'] = form
                    template = get_template("adm_areaspolideportivo/modal/formarea.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'editactividad':
                try:
                    data['id'] = request.GET['id']
                    data['filtro'] = filtro = ActividadPolideportivo.objects.get(pk=request.GET['id'])
                    form = ActividadPolideportivoForm(initial={'disciplina': filtro.disciplina,
                                                               'nombre': filtro.nombre,
                                                               'tipoactividad': filtro.tipoactividad,
                                                               'descripcion': filtro.descripcion,
                                                               'fechainicio': filtro.fechainicio,
                                                               'fechafin': filtro.fechafin,
                                                               'responsable': filtro.responsable,
                                                               'cupo': filtro.cupo,
                                                               'mostrar': filtro.mostrar,
                                                               'portada': filtro.portada,
                                                               'interno': filtro.interno,
                                                               'externo': filtro.externo,
                                                               'valor': filtro.valor,
                                                               'carrera': filtro.carreras.all(),
                                                               'numacompanantes': filtro.numacompanantes})
                    data['form'] = form
                    template = get_template("adm_areaspolideportivo/modal/formactividad.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'editturno':
                try:
                    data['id'] = request.GET['id']
                    data['filtro'] = filtro = TurnoPolideportivo.objects.get(pk=request.GET['id'])
                    form = TurnoPolideportivoForm(initial=model_to_dict(filtro))
                    data['form'] = form
                    template = get_template("adm_areaspolideportivo/modal/formturno.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'editsancion':
                try:
                    data['id'] = request.GET['id']
                    data['filtro'] = filtro = SancionPolideportivo.objects.get(pk=request.GET['id'])
                    form = SancionPolideportivoForm(initial=model_to_dict(filtro))
                    data['form'] = form
                    template = get_template("adm_areaspolideportivo/modal/formsancion.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'editpolitica':
                try:
                    data['id'] = request.GET['id']
                    data['filtro'] = filtro = PoliticaPolideportivo.objects.get(pk=request.GET['id'])
                    form = PoliticaPolideportivoForm(initial=model_to_dict(filtro))
                    data['form'] = form
                    template = get_template("adm_areaspolideportivo/modal/formpolitica.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'editdis':
                try:
                    data['id'] = request.GET['id']
                    data['filtro'] = filtro = DisciplinaDeportiva.objects.get(pk=request.GET['id'])
                    form = DisciplinaDeportivaForm(initial=model_to_dict(filtro))
                    data['form'] = form
                    template = get_template("adm_areaspolideportivo/modal/formdisciplina.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass
            elif action == 'editseccion':
                try:
                    data['id'] = request.GET['id']
                    data['filtro'] = filtro = SeccionPolideportivo.objects.get(pk=request.GET['id'])
                    form = SeccionPolideportivoForm(initial=model_to_dict(filtro))
                    data['form'] = form
                    template = get_template("adm_areaspolideportivo/modal/formseccion.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'edithorario':
                try:
                    data['id'] = request.GET['id']
                    data['filtro'] = filtro = HorarioActividadPolideportivo.objects.get(pk=request.GET['id'])
                    data['actividad'] = actividad = filtro.actividad
                    form = HorarioActividadPolideportivoForm(initial={'turno': filtro.turno,
                                                                      'mostrar': filtro.mostrar,
                                                                      'dia': filtro.dia,
                                                                      'instructor': filtro.instructor,
                                                                      'fechainicio': filtro.fechainicio,
                                                                      'fechafin': filtro.fechafin,
                                                                      })
                    form.cargar_instructores(actividad)
                    data['form'] = form
                    template = get_template("adm_areaspolideportivo/modal/formhorario.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'editimplementoact':
                try:
                    data['id'] = request.GET['id']
                    data['filtro'] = filtro = ImplementosActividad.objects.get(pk=request.GET['id'])
                    form = ImplementosActividadForm(initial=model_to_dict(filtro))
                    data['form'] = form
                    template = get_template("adm_areaspolideportivo/modal/formimplemento.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'editperfil':
                try:
                    data['id'] = request.GET['id']
                    data['filtro'] = filtro = PerfilesActividad.objects.get(pk=request.GET['id'])
                    form = PerfilesActividadForm(initial=model_to_dict(filtro))
                    data['form'] = form
                    template = get_template("adm_areaspolideportivo/modal/formperfil.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'editimplemento':
                try:
                    data['id'] = request.GET['id']
                    data['filtro'] = filtro = Utensilios.objects.get(pk=request.GET['id'])
                    form = ImplementoForm(initial=model_to_dict(filtro))
                    data['form'] = form
                    template = get_template("adm_areaspolideportivo/modal/formutensilio.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'editinstructoract':
                try:
                    data['id'] = id = request.GET['id']
                    data['idpadre'] = request.GET['idpadre']
                    filtro = InstructorActividadPolideportivo.objects.get(id=id)
                    form = InstructorActividadPolideportivoForm(initial=model_to_dict(filtro))
                    data['form2'] = form
                    template = get_template("adm_areaspolideportivo/modal/forminstructoract.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'verfotos':
                try:
                    data['id'] = request.GET['id']
                    data['filtro'] = filtro = AreaPolideportivo.objects.get(pk=request.GET['id'])
                    data['title'] = 'Área | Fotos'
                    data['listado'] = listado = filtro.fotosareapolideportivo_set.filter(status=True).order_by('orden')
                    return render(request, 'adm_areaspolideportivo/verfotos.html', data)
                except Exception as ex:
                    pass

            elif action == 'actividades':
                try:
                    data['id'] = id = encrypt_id(request.GET['id'])
                    area = AreaPolideportivo.objects.get(pk=id)
                    filtro, search, url_vars = Q(status=True, area_id=id), request.GET.get('s', ''), f'&action={action}&id={encrypt(id)}'
                    if search:
                        data['search'] = search
                        filtro = filtro & (Q(nombre__icontains=search))
                        url_vars += '&s={}'.format(search)
                    listado = ActividadPolideportivo.objects.filter(filtro).order_by('id')
                    paging = MiPaginador(listado, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['url_vars'] = url_vars
                    data['area'] = area.nombre
                    data['title'] = 'Área | Actividades'
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['listado'] = page.object_list
                    request.session['viewactivo'] = 1
                    return render(request, 'adm_areaspolideportivo/viewactividad.html', data)
                except Exception as ex:
                    pass

            elif action == 'turnos':
                try:
                    data['title'] = 'Turnos'
                    search, filtros, url_vars = request.GET.get('s', ''), Q(status=True), f'&action={action}'
                    # data['listado'] = TurnoPolideportivo.objects.filter(status=True).order_by('turno')
                    # url_vars += '?action={}'.format(action)
                    if search:
                        data['search'] = search
                        s = search.split()
                        if len(s) == 1:
                            filtros = filtros & (Q(comienza__icontains=search) |
                                                 Q(termina__icontains=search))
                        else:
                            filtros = filtros & (Q(comienza__icontains=s[0]) &
                                                 Q(termina__icontains=s[1]))
                        url_vars += '&search={}'.format(search)
                    turnos = TurnoPolideportivo.objects.filter(filtros).order_by('comienza')
                    paging = MiPaginador(turnos, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['url_vars'] = url_vars
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['listado'] = page.object_list
                    request.session['viewactivo'] = 2
                    return render(request, 'adm_areaspolideportivo/viewturno.html', data)
                except Exception as ex:
                    pass

            elif action == 'disciplina':
                try:
                    data['title'] = 'Disciplinas'
                    filtro, search, url_vars = Q(status=True), request.GET.get('s', ''), f'&action={action}'
                    if search:
                        data['search'] = search
                        filtro = filtro & (Q(descripcion__icontains=search))
                        url_vars += '&s={}'.format(search)
                    disciplina = DisciplinaDeportiva.objects.filter(filtro).order_by('descripcion')
                    paging = MiPaginador(disciplina, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['url_vars'] = url_vars
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['listado'] = page.object_list
                    request.session['viewactivo'] = 3
                    return render(request, 'adm_areaspolideportivo/viewdisciplina.html', data)
                except Exception as ex:
                    pass

            elif action == 'implemento':
                try:
                    data['title'] = 'Gestión de implementos disponibles'
                    filtro, search, url_vars = Q(status=True), request.GET.get('s', ''), f'&action={action}'
                    if search:
                        data['search'] = search
                        filtro = filtro & (Q(descripcion__icontains=search))
                        url_vars += '&s={}'.format(search)
                    implementos = Utensilios.objects.filter(filtro)
                    paging = MiPaginador(implementos, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['url_vars'] = url_vars
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['listado'] = page.object_list
                    request.session['viewactivo'] = 4
                    return render(request, 'adm_areaspolideportivo/viewutensilio.html', data)
                except Exception as ex:
                    pass

            elif action == 'perfiles':
                try:
                    data['filtro'] = filtro = ActividadPolideportivo.objects.get(pk=request.GET['id'])
                    data['title'] = 'Gestión de perfiles de %s' % (filtro.nombre)
                    data['listado'] = PerfilesActividad.objects.filter(actividad=filtro, status=True).order_by('id')
                    return render(request, 'adm_areaspolideportivo/viewperfil.html', data)
                except Exception as ex:
                    pass

            elif action == 'sancion':
                try:
                    data['title'] = 'Gestión de sanciones'
                    filtro, search, url_vars = Q(status=True), request.GET.get('s', ''), f'&action={action}'
                    if search:
                        data['search'] = search
                        filtro = filtro & (Q(descripcion__icontains=search))
                        url_vars += '&s={}'.format(search)
                    sanciones = SancionPolideportivo.objects.filter(filtro).order_by('-pk')
                    paging = MiPaginador(sanciones, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['url_vars'] = url_vars
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['listado'] = page.object_list
                    request.session['viewactivo'] = 5
                    return render(request, 'adm_areaspolideportivo/viewsancion.html', data)
                except Exception as ex:
                    pass

            elif action == 'politica':
                try:
                    data['title'] = 'Gestión de políticas'
                    filtro, search, url_vars = Q(status=True), request.GET.get('s', ''), f'&action={action}'
                    if search:
                        data['search'] = search
                        filtro = filtro & (Q(descripcion__icontains=search) | Q(nombre__icontains=search))
                        url_vars += '&s={}'.format(search)
                    politicas = PoliticaPolideportivo.objects.filter(filtro).order_by('pk')
                    paging = MiPaginador(politicas, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['url_vars'] = url_vars
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['listado'] = page.object_list
                    request.session['viewactivo'] = 6
                    return render(request, 'adm_areaspolideportivo/viewpolitica.html', data)
                except Exception as ex:
                    pass

            elif action == 'secciones':
                try:
                    data['id'] = request.GET['id']
                    data['filtro'] = filtro = AreaPolideportivo.objects.get(pk=request.GET['id'])
                    data['title'] = 'Secciones de {}'.format(filtro.nombre)
                    data['listado'] = SeccionPolideportivo.objects.filter(area_id=request.GET['id'], status=True)
                    return render(request, 'adm_areaspolideportivo/viewseccion.html', data)
                except Exception as ex:
                    pass

            elif action == 'implementosact':
                try:
                    data['title'] = 'Actividad | Implementos'
                    data['id'] = id = int(encrypt(request.GET['id']))
                    filtro, search, url_vars = Q(actividad_id=id, status=True), request.GET.get('s', ''), f'&action={action}'
                    if search:
                        data['search'] = search
                        filtro = filtro & (Q(utensilio__descripcion__icontains=search))
                        url_vars += '&s={}'.format(search)
                    listado = ImplementosActividad.objects.filter(filtro)
                    paging = MiPaginador(listado, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['url_vars'] = url_vars
                    data['filtro'] = ActividadPolideportivo.objects.get(pk=id)
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['listado'] = page.object_list
                    return render(request, 'adm_areaspolideportivo/viewimplemento.html', data)
                except Exception as ex:
                    pass

            elif action == 'instructoresactividad':
                try:
                    data['title'] = 'Actividad | Instructores'
                    data['id'] = id = int(encrypt(request.GET['id']))
                    filtro, search, url_vars = Q(status=True, actividad_id=id), request.GET.get('s', ''), f'&action={action}'
                    if search:
                        data['search'] = search
                        filtro = filtro & (Q(instructor__persona__apellido1__icontains=search) |
                                           Q(instructor__persona__nombres__icontains=search) |
                                           Q(instructor__persona__cedula__icontains=search))
                        url_vars += '&s={}'.format(search)
                    listado = InstructorActividadPolideportivo.objects.filter(filtro).order_by('-id')
                    paging = MiPaginador(listado, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['url_vars'] = url_vars
                    data['filtro'] = ActividadPolideportivo.objects.get(pk=id)
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['listado'] = page.object_list
                    request.session['viewactivo'] = 1
                    return render(request, 'adm_areaspolideportivo/viewinstructoractividad.html', data)
                except Exception as ex:
                    pass

            elif action == 'instructores':
                try:
                    data['title'] = u'Gestión de instructores'
                    search, filtro, url_vars = request.GET.get('s', ''), Q(status=True), f'&action={action}'
                    if search:
                        filtro = filtro & Q(nombre__icontains=search)
                        url_vars += '&s=' + search
                        data['search'] = search
                    listado = InstructorPolideportivo.objects.filter(filtro).order_by('-id')
                    paging = MiPaginador(listado, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data["url_vars"] = url_vars
                    data['listado'] = page.object_list
                    data['totcount'] = listado.count()
                    data['email_domain'] = EMAIL_DOMAIN
                    request.session['viewactivo'] = 7
                    return render(request, 'adm_areaspolideportivo/viewinstructor.html', data)
                except Exception as ex:
                    pass

            elif action == 'addfoto':
                try:
                    data['id'] = request.GET['id']
                    data['filtro'] = filtro = AreaPolideportivo.objects.get(pk=request.GET['id'])
                    form = FotoAreaPolideportivoForm()
                    data['form'] = form
                    template = get_template("adm_areaspolideportivo/modal/formfotoarea.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'editfoto':
                try:
                    data['id'] = request.GET['id']
                    data['filtro'] = filtro = FotosAreaPolideportivo.objects.get(pk=request.GET['id'])
                    form = FotoAreaPolideportivoForm(initial=model_to_dict(filtro))
                    data['form'] = form
                    template = get_template("adm_areaspolideportivo/modal/formfotoarea.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'horarios':
                try:
                    data['title'] = u'Actividad | Horarios'
                    id = encrypt_id(request.GET['id'])
                    idp = request.GET.get('idp', '')
                    data['semana'] = ['Lunes', 'Martes', 'Miercoles', 'Jueves', 'Viernes', 'Sabado', 'Domingo']
                    data['turnos'] = TurnoPolideportivo.objects.filter(status=True).order_by('comienza','termina')
                    data['actividad'] = ActividadPolideportivo.objects.get(pk=id)
                    data['planificacion'] = PlanificacionActividad.objects.get(pk=encrypt_id(idp)) if idp else None
                    return render(request, "adm_areaspolideportivo/viewhorario.html", data)
                except Exception as ex:
                    pass

            elif action == 'buscaresponsable':
                try:
                    q = request.GET['q'].upper().strip()
                    s = q.split(" ")
                    personas = DistributivoPersona.objects.filter(status=True).order_by('persona_id').distinct('persona_id')
                    if len(s) == 1:
                        per = personas.filter((Q(persona__nombres__icontains=q) | Q(persona__apellido1__icontains=q) | Q(
                            persona__apellido2__icontains=q) | Q(persona__cedula__contains=q))).distinct()[:15]
                    elif len(s) == 2:
                        per = personas.filter((Q(persona__apellido1__contains=s[0]) & Q(persona__apellido2__contains=s[1])) | (
                                Q(nombres__icontains=s[0]) & Q(
                            persona__nombres__icontains=s[1])) | (
                                                      Q(persona__nombres__icontains=s[0]) & Q(
                                                  persona__apellido1__contains=s[1]))).distinct()[:15]
                    else:
                        per = personas.filter((Q(persona__nombres__contains=s[0]) & Q(persona__apellido1__contains=s[1]) & Q(
                            persona__apellido2__contains=s[2])) | (Q(persona__nombres__contains=s[0]) & Q(
                            persona__nombres__contains=s[1]) & Q(persona__apellido1__contains=s[2]))).distinct()[:15]

                    data = {"result": "ok",
                            "results": [{"id": x.persona.id, "name": str(x.persona.nombre_completo())}
                                        for x in per]}
                    return JsonResponse(data)
                except Exception as ex:
                    pass

            elif action == 'buscapersona':
                try:
                    q = request.GET['q'].upper().strip()
                    s = q.split(" ")
                    if len(s) == 1:
                        per = Persona.objects.filter((Q(nombres__icontains=q) | Q(apellido1__icontains=q) | Q(
                            apellido2__icontains=q) | Q(cedula__contains=q))).distinct()[:15]
                    elif len(s) == 2:
                        per = Persona.objects.filter((Q(apellido1__contains=s[0]) & Q(apellido2__contains=s[1])) | (
                                Q(nombres__icontains=s[0]) & Q(
                            nombres__icontains=s[1])) | (
                                                             Q(nombres__icontains=s[0]) & Q(
                                                         apellido1__contains=s[1]))).distinct()[:15]
                    else:
                        per = Persona.objects.filter((Q(nombres__contains=s[0]) & Q(apellido1__contains=s[1]) & Q(
                            apellido2__contains=s[2])) | (Q(nombres__contains=s[0]) & Q(
                            nombres__contains=s[1]) & Q(apellido1__contains=s[2]))).distinct()[:15]

                    data = {"result": "ok",
                            "results": [{"id": x.id, "name": str(x.nombre_completo())}
                                        for x in per]}
                    return JsonResponse(data)
                except Exception as ex:
                    pass

            elif action == 'buscarpersonas':
                try:
                    resp = filtro_persona_select(request)
                    return HttpResponse(json.dumps({'status': True, 'results': resp}))
                except Exception as ex:
                    pass

            elif action == 'buscainstructor':
                try:
                    q = request.GET['q'].upper().strip()
                    s = q.split(" ")
                    if len(s) == 1:
                        per = InstructorPolideportivo.objects.filter((Q(persona__nombres__icontains=q) | Q(persona__apellido1__icontains=q) | Q(
                            persona__apellido2__icontains=q) | Q(persona__cedula__contains=q)), activo=True).distinct()[:15]
                    elif len(s) == 2:
                        per = InstructorPolideportivo.objects.filter((Q(persona__apellido1__contains=s[0]) & Q(persona__apellido2__contains=s[1])) | (
                                Q(persona__nombres__icontains=s[0]) & Q(
                            persona__nombres__icontains=s[1])) | (
                                                                             Q(persona__nombres__icontains=s[0]) & Q(
                                                                         persona__apellido1__contains=s[1])), activo=True).distinct()[:15]
                    else:
                        per = InstructorPolideportivo.objects.filter((Q(persona__nombres__contains=s[0]) & Q(persona__apellido1__contains=s[1]) & Q(
                            persona__apellido2__contains=s[2])) | (Q(persona__nombres__contains=s[0]) & Q(
                            persona__nombres__contains=s[1]) & Q(persona__apellido1__contains=s[2])), activo=True).distinct()[:15]

                    data = {"result": "ok",
                            "results": [{"id": x.id, "name": str(x.persona)}
                                        for x in per]}
                    return JsonResponse(data)
                except Exception as ex:
                    pass

            elif action == 'reporteria':
                try:
                    area, disciplina, actividad, filtro = request.GET.get('area', ''), \
                                                          request.GET.get('disciplina', ''), \
                                                          request.GET.get('actividad', ''), \
                                                          Q(status=True)
                    data['title'] = 'Reportes de actividades'
                    if area or disciplina:
                        if area and area != '0':
                            filtro = filtro & Q(area_id=int(area))
                        elif disciplina and disciplina != '0':
                            filtro = filtro & Q(disciplina_id=int(disciplina))

                        wb = openxl.Workbook()
                        wb["Sheet"].title = "Reporte_actividades"
                        ws = wb.active
                        style_title = openxlFont(name='Arial', size=16, bold=True)
                        style_cab = openxlFont(name='Arial', size=10, bold=True)
                        alinear = alin(horizontal="center", vertical="center")
                        response = HttpResponse(content_type="application/ms-excel")
                        response[
                            'Content-Disposition'] = 'attachment; filename=Reporte de actividades' + '-' + random.randint(
                            1, 10000).__str__() + '.xlsx'
                        ws.column_dimensions['A'].width = 5
                        ws.column_dimensions['B'].width = 25
                        ws.column_dimensions['C'].width = 5
                        ws.column_dimensions['D'].width = 30
                        ws.column_dimensions['E'].width = 20
                        ws.column_dimensions['F'].width = 30
                        ws.column_dimensions['G'].width = 15
                        ws.column_dimensions['H'].width = 15
                        ws.column_dimensions['I'].width = 10
                        ws.column_dimensions['J'].width = 10
                        ws.merge_cells('A1:M1')
                        ws['A1'] = 'REPORTE DE ACTIVIDADES EN ESPACIOS DE POLIDEPORTIVO'
                        celda1 = ws['A1']
                        celda1.font = style_title
                        celda1.alignment = alinear
                        ws['A4'].alignment = alinear
                        ws['B4'].alignment = alinear
                        columns = [u"N°", u"Área", 'Cantidad', u"Actividad", u"Diciplína",
                                   u"Instructor", u"Fecha inicio", u"Fecha fin",
                                   u"Visible", u"¿Encuesta configurada?", u"Enc. Vigente"
                                   ]
                        row_num = 3
                        for col_num in range(0, len(columns)):
                            celda = ws.cell(row=row_num, column=(col_num + 1), value=columns[col_num])
                            celda.font = style_cab

                        mensaje = 'NO REGISTRA'
                        row_num = 4
                        numero = 0
                        if area and area != '0' or disciplina and disciplina != '0':
                            listado = ActividadPolideportivo.objects.filter(filtro)
                            for actividad in listado:
                                numero += 1
                                ws.cell(row=row_num, column=1, value=numero)
                                ws.cell(row=row_num, column=2, value=str(actividad.area))
                                ws.cell(row=row_num, column=3, value=str(len(listado)))
                                ws.cell(row=row_num, column=4, value=str(actividad))
                                ws.cell(row=row_num, column=5, value=str(actividad.disciplina))
                                ws.cell(row=row_num, column=6, value=str(actividad.responsable.nombre_completo_minus()))
                                ws.cell(row=row_num, column=7, value=str(actividad.fechainicio))
                                ws.cell(row=row_num, column=8, value=str(actividad.fechafin))
                                ws.cell(row=row_num, column=9, value="Si" if actividad.mostrar else "No")
                                ws.cell(row=row_num, column=10, value="Si" if actividad.encuesta_configurada() else "No")
                                ws.cell(row=row_num, column=11, value="Si" if actividad.encuesta() else "No")
                                row_num += 1
                        else:
                            areas = AreaPolideportivo.objects.filter(status=True)
                            for idx, area in enumerate(areas, start=1):
                                listado = area.actividades()
                                filafinal = row_num + len(listado) - 1
                                ws.merge_cells(f'A{row_num}:A{filafinal}')
                                ws.merge_cells(f'B{row_num}:B{filafinal}')
                                ws.merge_cells(f'C{row_num}:C{filafinal}')
                                ws.cell(row=row_num, column=1, value=idx).alignment = alinear
                                ws.cell(row=row_num, column=2, value=str(area)).alignment = alinear
                                ws.cell(row=row_num, column=3, value=len(listado)).alignment = alinear
                                for actividad in listado:
                                    ws.cell(row=row_num, column=4, value=str(actividad))
                                    ws.cell(row=row_num, column=5, value=str(actividad.disciplina))
                                    ws.cell(row=row_num, column=6, value=str(actividad.responsable.nombre_completo_minus()))
                                    ws.cell(row=row_num, column=7, value=str(actividad.fechainicio))
                                    ws.cell(row=row_num, column=8, value=str(actividad.fechafin))
                                    ws.cell(row=row_num, column=9, value="Si" if actividad.mostrar else "No")
                                    ws.cell(row=row_num, column=10, value="Si" if actividad.encuesta_configurada() else "No")
                                    ws.cell(row=row_num, column=11, value="Si" if actividad.encuesta() else "No")
                                    row_num += 1
                        wb.save(response)
                        return response
                    elif actividad:
                        filtro = filtro & Q(actividad_id=int(actividad))
                        wb = openxl.Workbook()
                        wb["Sheet"].title = "Reporte_horarios"
                        ws = wb.active
                        style_title = openxlFont(name='Arial', size=16, bold=True)
                        style_cab = openxlFont(name='Arial', size=10, bold=True)
                        alinear = alin(horizontal="center", vertical="center")
                        response = HttpResponse(content_type="application/ms-excel")
                        response[
                            'Content-Disposition'] = 'attachment; filename=Reporte de horarios' + '-' + random.randint(
                            1, 10000).__str__() + '.xlsx'
                        ws.column_dimensions['B'].width = 25
                        ws.column_dimensions['C'].width = 40
                        ws.column_dimensions['D'].width = 20
                        ws.column_dimensions['E'].width = 30
                        ws.column_dimensions['F'].width = 25
                        ws.column_dimensions['G'].width = 20
                        ws.merge_cells('A1:M1')
                        ws['A1'] = 'REPORTE DE HORARIOS POLIDEPORTIVO'
                        celda1 = ws['A1']
                        celda1.font = style_title
                        celda1.alignment = alinear

                        columns = [u"N°", u"ACTIVIDAD", u"HORARIO", "DIA",
                                   u"F.INICIO", u"F.FIN", "INSTRUCTOR"
                                   ]
                        row_num = 3
                        for col_num in range(0, len(columns)):
                            celda = ws.cell(row=row_num, column=(col_num + 1), value=columns[col_num])
                            celda.font = style_cab

                        mensaje = 'NO REGISTRA'
                        row_num = 4
                        numero = 0
                        listado = HorarioActividadPolideportivo.objects.filter(filtro).order_by('dia')
                        for list in listado:
                            numero += 1
                            ws.cell(row=row_num, column=1, value=numero)
                            ws.cell(row=row_num, column=2, value=str(list.actividad))
                            ws.cell(row=row_num, column=3, value=str(list.turno))
                            ws.cell(row=row_num, column=4, value=str(list.get_dia_display()))
                            ws.cell(row=row_num, column=5, value=str(list.fechainicio))
                            ws.cell(row=row_num, column=6, value=str(list.fechafin))
                            ws.cell(row=row_num, column=7, value=str(list.instructor))
                            row_num += 1
                        wb.save(response)
                        return response
                    data['areas'] = AreaPolideportivo.objects.filter(filtro).order_by('-id')
                    data['disciplinas'] = DisciplinaDeportiva.objects.filter(filtro).order_by('-id')
                    data['actividades'] = ActividadPolideportivo.objects.filter(filtro).order_by('-id')
                    template = get_template('adm_areaspolideportivo/modal/reportes.html')
                    return JsonResponse({'result': True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'clubes':
                data['title'] = 'Clubes'
                search, filtro, url_vars = request.GET.get('s', ''), Q(status=True), f'&action={action}'
                if search:
                    data['search'] = search.strip()
                    filtro = filtro & Q(nombre__icontains=search)
                    url_vars += '&s=' + search
                listado = ClubPoli.objects.filter(filtro).order_by('-id')
                paging = MiPaginador(listado, 20)
                p = 1
                try:
                    paginasesion = 1
                    if 'paginador' in request.session:
                        paginasesion = int(request.session['paginador'])
                    if 'page' in request.GET:
                        p = int(request.GET['page'])
                    else:
                        p = paginasesion
                    try:
                        page = paging.page(p)
                    except:
                        p = 1
                    page = paging.page(p)
                except:
                    page = paging.page(p)
                request.session['paginador'] = p
                data['paging'] = paging
                data['rangospaging'] = paging.rangos_paginado(p)
                data['page'] = page
                data["url_vars"] = url_vars
                data['listado'] = page.object_list
                request.session['viewactivo'] = 8
                return render(request, 'adm_areaspolideportivo/viewclub.html', data)

            elif action == 'addclub':
                try:
                    form = ClubPoliForm()
                    data['form'] = form
                    template = get_template("adm_areaspolideportivo/modal/formclubpoli.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'editclub':
                try:
                    id = int(encrypt(request.GET['id']))
                    data['club'] = club = ClubPoli.objects.get(id=id)
                    form = ClubPoliForm(initial=model_to_dict(club))
                    data['form'] = form
                    template = get_template("adm_areaspolideportivo/modal/formclubpoli.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'integrantes':
                try:
                    data['title'] = 'Club | Integrantes'
                    id = int(encrypt(request.GET['id']))
                    search, filtro, url_vars, tipo = request.GET.get('s', ''), Q(status=True, club_id=id), f'&action={action}', request.GET.get('tipo', '')

                    if tipo:
                        data['t'] = t = int(tipo)
                        filtro = filtro & Q(tipointegrante=t)
                        url_vars += '&t=' + tipo
                    if search:
                        data['search'] = search.strip()
                        s = search.split()
                        if s == 1:
                            filtro = filtro & Q(integante__cedula__icontains=search)
                        elif s == 2:
                            filtro = filtro & (Q(integante__apellido1__icontains=s[0]) |
                                               Q(integante__apellido2__icontains=s[1]))

                        url_vars += '&s=' + search
                    listado = IntegranteClubPoli.objects.filter(filtro).order_by('-id')
                    paging = MiPaginador(listado, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['filtro'] = ClubPoli.objects.get(id=id)
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['tipos'] = TIPO_INTEGRANTE
                    data['page'] = page
                    data["url_vars"] = url_vars
                    data['listado'] = page.object_list
                    request.session['viewactivo'] = 8
                    return render(request, 'adm_areaspolideportivo/viewintegrantes.html', data)
                except Exception as ex:
                    print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
                    pass

            elif action == 'addintegrante':
                try:
                    form = IntegranteClubForm()
                    data['id'] = int(encrypt(request.GET['idpadre']))
                    data['form'] = form
                    template = get_template("adm_areaspolideportivo/modal/formintegrante.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'descuentos':
                # Requisitos
                try:
                    form = DescuentoValorActividadForm()
                    data['filtro'] = perfiles = PerfilesActividad.objects.get(pk=int(encrypt(request.GET['id'])))
                    data['form'] = form
                    template = get_template("adm_areaspolideportivo/modal/formdescuentos.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'addencuesta':
                try:
                    data['id'] = id = encrypt_id(request.GET['id'])
                    actividad = ActividadPolideportivo.objects.get(id=id)
                    content_type = ContentType.objects.get_for_model(actividad)
                    encuesta = EncuestaProceso.objects.filter(object_id=actividad.id, content_type=content_type, status=True).first()
                    if encuesta:
                        form = EncuestaPreguntaForm(initial=model_to_dict(encuesta))
                        data['preguntas'] = PreguntaEncuestaProceso.objects.filter(status=True, encuesta=encuesta)
                    else:
                        form = EncuestaPreguntaForm()
                    data['form'] = form
                    data['switchery'] = True
                    template = get_template('formencuesta.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            # Sitio Web Unemi Deporte
            elif action == 'sitioweb':
                try:
                    data['title'] = u'Unemi Deporte'
                    seccionactiva = request.GET.get('s_activa', 1)
                    titulos = TituloWebSite.objects.filter(status=True)
                    if len(SECCION) > len(titulos):
                        for s in SECCION:
                            titulo = TituloWebSite.objects.filter(status=True, seccion=s[0])
                            if not titulo.exists():
                                title = ''
                                subtitle = ''
                                if s[0] == 1:
                                    title = 'Nuestras Escuelas Formativas'
                                    subtitle = '¡Conoce nuestras Escuelas Formativas!'
                                elif s[0] == 2:
                                    title = 'Nuestros Espacios Deportivos'
                                    subtitle = '¡Conoce nuestros Espacios Deportivos!'
                                elif s[0] == 3:
                                    title = 'Nuestros Vacacionales'
                                    subtitle = '¡Conoce nuestros Vacacionales!'
                                elif s[0] == 4:
                                    title = 'Noticias Deportivas'
                                    subtitle = '¡Conoce todas las noticias ocurridas en Unemi Deporte!'
                                elif s[0] == 5:
                                    title = 'Nuestros Logros'
                                    subtitle = '¡Conoce nuestros Logros!'
                                elif s[0] == 6:
                                    title = 'Nuestro Equipo'
                                    subtitle = '¡Conoce a nuestro Instructores de las diferentes actividades que disponemos!'
                                titulo = TituloWebSite(seccion=s[0], titulo=title, subtitulo=subtitle)
                                titulo.save(request)
                    data['secciones'] = SECCION
                    data['s_activa'] = int(seccionactiva)
                    data['titulos'] = TituloWebSite.objects.filter(status=True).order_by('seccion')
                    request.session['viewactivo'] = 9
                    return render(request, 'adm_areaspolideportivo/unemideporte.html', data)
                except Exception as ex:
                    messages.error(f'{ex}')
                    print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))

            elif action == 'editsecciontitle':
                try:
                    data['id'] = id = int(encrypt(request.GET['id']))
                    titulo = TituloWebSite.objects.get(id=id)
                    data['switchery'] = True
                    form = TituloWebSiteForm(initial=model_to_dict(titulo), instancia=titulo)
                    data['form'] = form
                    template = get_template("ajaxformmodal.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'mensaje': f'Error: {ex}'})

            elif action == 'addcuerpo':
                try:
                    data['idp'] = encrypt_id(request.GET['id'])
                    form = CuerpoWebSiteForm()
                    form.fields['ubicacion'].initial = request.GET['idex']
                    data['form'] = form
                    data['switchery'] = True
                    template = get_template('ajaxformmodal.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'editcuerpo':
                try:
                    data['id'] = id = encrypt_id(request.GET['id'])
                    cuerpo = CuerpoWebSite.objects.get(id=id)
                    form = CuerpoWebSiteForm(instancia=cuerpo, initial=model_to_dict(cuerpo))
                    data['form'] = form
                    data['switchery'] = True
                    template = get_template('ajaxformmodal.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')


            # NOTICIAS
            elif action == 'noticias':
                data['title'] = 'Noticias'
                search, filtro, url_vars = request.GET.get('s', ''), Q(status=True), f'&action={action}'
                if search:
                    data['search'] = search.strip()
                    filtro = filtro & Q(titulo__icontains=search) | Q(subtitulo__icontains=search)
                    url_vars += '&s=' + search
                listado = NoticiaDeportiva.objects.filter(filtro).order_by('-principal', '-fecha_creacion')
                paging = MiPaginador(listado, 20)
                p = 1
                try:
                    paginasesion = 1
                    if 'paginador' in request.session:
                        paginasesion = int(request.session['paginador'])
                    if 'page' in request.GET:
                        p = int(request.GET['page'])
                    else:
                        p = paginasesion
                    try:
                        page = paging.page(p)
                    except:
                        p = 1
                    page = paging.page(p)
                except:
                    page = paging.page(p)
                request.session['paginador'] = p
                data['paging'] = paging
                data['rangospaging'] = paging.rangos_paginado(p)
                data['page'] = page
                data["url_vars"] = url_vars
                data['listado'] = page.object_list
                request.session['viewactivo'] = 10
                return render(request, 'adm_areaspolideportivo/noticias.html', data)

            elif action == 'addnoticia':
                try:
                    form = NoticiaDeportivaForm()
                    visible_fields = form.visible_fields()
                    total_fields = len(visible_fields)
                    lista = [(1, 'Información Básica', visible_fields[:5]),
                             (2, 'Contenido', visible_fields[5:total_fields])
                             ]
                    data['form'] = lista
                    data['switchery'] = True
                    template = get_template('adm_areaspolideportivo/modal/formnoticia.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'editnoticia':
                try:
                    data['id'] = id = encrypt_id(request.GET['id'])
                    noticia = NoticiaDeportiva.objects.get(id=id)
                    form = NoticiaDeportivaForm(instancia=noticia, initial=model_to_dict(noticia))
                    visible_fields = form.visible_fields()
                    total_fields = len(visible_fields)
                    lista = [(1, 'Información Básica', visible_fields[:5]),
                             (2, 'Contenido', visible_fields[5:total_fields])
                             ]
                    data['form'] = lista
                    data['switchery'] = True
                    template = get_template('adm_areaspolideportivo/modal/formnoticia.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'previsualizarnoticia':
                try:
                    data['id'] = id = encrypt_id(request.GET['id'])
                    data['noticia'] = NoticiaDeportiva.objects.get(id=id)
                    template = get_template('adm_areaspolideportivo/modal/previsualizarnoticia.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            # PLANIFICACIONES
            elif action == 'planificaciones':
                try:
                    id = encrypt_id(request.GET['id'])
                    data['actividad'] = actividad = ActividadPolideportivo.objects.get(pk=id)
                    filtro, search, url_vars = Q(status=True, actividad_id=id), request.GET.get('s', ''), f'&action={action}'
                    if search:
                        data['search'] = search
                        filtro = filtro & (Q(nombre__icontains=search))
                        url_vars += '&s={}'.format(search)
                    listado = PlanificacionActividad.objects.filter(filtro).order_by('id')
                    paging = MiPaginador(listado, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['url_vars'] = url_vars
                    data['title'] = 'Actvidad | Planificaciones'
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['listado'] = page.object_list
                    request.session['viewactivo'] = 1
                    return render(request, 'adm_areaspolideportivo/planificacion.html', data)
                except Exception as ex:
                    pass

            elif action == 'addplanificacion':
                try:
                    form = PlanificacionActividadForm()
                    data['idp'] = encrypt_id(request.GET['id'])
                    data['form'] = form
                    data['switchery'] = True
                    template = get_template('ajaxformmodal.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'editplanificacion':
                try:
                    data['id'] = id = encrypt_id(request.GET['id'])
                    instancia = PlanificacionActividad.objects.get(id=id)
                    form = PlanificacionActividadForm(instancia=instancia, initial=model_to_dict(instancia))
                    data['form'] = form
                    data['switchery'] = True
                    template = get_template('ajaxformmodal.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            return HttpResponseRedirect(request.path)
        else:
            try:
                data['title'] = u'Áreas'
                search, filtro, url_vars = request.GET.get('s', ''), Q(status=True), ''
                if search:
                    filtro = filtro & Q(nombre__icontains=search)
                    url_vars += '&s=' + search
                    data['search'] = search
                listado = AreaPolideportivo.objects.filter(filtro).order_by('-id')
                paging = MiPaginador(listado, 20)
                p = 1
                try:
                    paginasesion = 1
                    if 'paginador' in request.session:
                        paginasesion = int(request.session['paginador'])
                    if 'page' in request.GET:
                        p = int(request.GET['page'])
                    else:
                        p = paginasesion
                    try:
                        page = paging.page(p)
                    except:
                        p = 1
                    page = paging.page(p)
                except:
                    page = paging.page(p)
                request.session['paginador'] = p
                data['paging'] = paging
                data['rangospaging'] = paging.rangos_paginado(p)
                data['page'] = page
                data["url_vars"] = url_vars
                data['listado'] = page.object_list
                data['totcount'] = listado.count()
                data['email_domain'] = EMAIL_DOMAIN
                request.session['viewactivo'] = 1
                return render(request, 'adm_areaspolideportivo/view.html', data)
            except Exception as ex:
                print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
                pass
