# -*- coding: UTF-8 -*-
import random
import sys

import xlsxwriter
import io
import openpyxl
import xlrd as xlrd
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import Group
from django.core.exceptions import ObjectDoesNotExist
from django.db import transaction, connection,connections
from django.contrib import messages
from xml.dom import minidom
from django.db.models.query_utils import Q
from django.http import HttpResponseRedirect, JsonResponse, HttpResponse
from django.shortcuts import render
from xlwt import Workbook
from xlwt.compat import xrange
from decimal import Decimal
from django.template.loader import get_template
from googletrans import Translator
from django.template import Context
from sga.excelbackground import actualizar_visible_horario_masivo, limpiar_cache_masivo
from bd.models import LogEntryLogin
from mobile.views import make_thumb_fotopersona, make_thumb_picture
from decorators import secure_module, last_access
from sga.commonviews import adduserdata, obtener_reporte, ficha_socioeconomica
from django.db.models import Count, PROTECT, Sum, Avg, Min, Max
from sga.forms import InscripcionForm, RecordAcademicoForm, HistoricoRecordAcademicoForm, CargarFotoForm, \
    DocumentoInscripcionForm, ConvalidacionInscripcionForm, \
    CambiomallaForm, HomologacionInscripcionForm, ConsiderarForm, NuevaInscripcionForm, CambioGrupoForm, \
    InscripcionTipoInscripcionForm, CambionivelmallaForm, \
    FechaInicioConvalidacionInscripcionForm, RetiradoMateriaForm, ImportarArchivoXLSForm, \
    FechaInicioPrimerNivelInscripcionForm, ActividadSakaiForm, ImportarListadoAlumnoForm, \
    ImportarListadoAlumnoSoporteForm, AsignarSoporteEstudiante, PersonaPPLForm, PersonaTituloUniversidadForm, \
    MatriculaSedeExamenForm, MallaHistoricaForm
from settings import DEFAULT_PASSWORD, ALUMNOS_GROUP_ID, UTILIZA_GRUPOS_ALUMNOS, EMAIL_DOMAIN, ARCHIVO_TIPO_GENERAL, \
    EMAIL_INSTITUCIONAL_AUTOMATICO, PREGUNTAS_INSCRIPCION, CORREO_OBLIGATORIO, GENERAR_TUMBAIL, \
    USA_TIPOS_INSCRIPCIONES, TIPO_INSCRIPCION_INICIAL, CONTROL_UNICO_CREDENCIALES, \
    MATRICULACION_LIBRE, TIPO_PERIODO_PROPEDEUTICO, TIPO_PERIODO_REGULAR, PROFESORES_GROUP_ID, MEDIA_URL
from sga.funciones import log, lista_correo, calculate_username, generar_usuario, \
    generar_nombre, resetear_clave, puede_realizar_accion, puede_modificar_inscripcion, MiPaginador, variable_valor, \
    querymysqlsakai, convertir_fecha, null_to_decimal, resetear_clave_admision_manual, resetear_clave_admision_ldap, \
    querymysqlconsulta, querymysqlmoodle, convertir_fecha_invertida, convertir_lista, ok_json, bad_json, elimina_tildes
from sga.funcionesxhtml2pdf import conviert_html_to_pdf
from sga.models import Persona, PersonaReligion , PersonaDocumentoPersonal, Inscripcion, RecordAcademico, DocumentosDeInscripcion, HistoricoRecordAcademico, \
    FotoPersona, Archivo, Grupo, ConvalidacionInscripcion, NivelMalla, EjeFormativo, AsignaturaMalla, Carrera, \
    SeguimientoEstudiante, miinstitucion, PreInscrito, InscripcionTipoInscripcion, Asignatura, Nivel, Matricula, Clase, \
    Sesion, RetiroCarrera, Modalidad, Sede, Administrativo, Profesor, TiempoDedicacionDocente, Coordinacion, \
    CUENTAS_CORREOS, Pais, Materia, MateriaAsignada, ProfesorMateria, PorcentajeCumplimientoAdmisionVirtual, \
    ActividadesSakaiAlumno, HorarioExamen, AuditoriaNotas, actualizar_nota_planificacion, Periodo, Provincia, Canton, \
    VirtualSoporteUsuario, VirtualSoporteUsuarioInscripcion, VirtualSoporteUsuarioProfesor, VirtualSoporteAsignado, \
    EvaluacionGenerica, LogEntryBackup, LogEntryBackupdos, AgregacionEliminacionMaterias, HistorialPersonaPPL, \
    PersonaTituloUniversidad, PlanificacionMateria, Reporte, Notificacion
from sga.models import InscripcionMalla, HomologacionInscripcion, InscripcionTesDrive, AutorprogramaAnalitico
from sga.reportes import run_report_v1
from sga.tasks import send_html_mail, conectar_cuenta
from datetime import datetime, timedelta
from sga.templatetags.sga_extras import encrypt
from django.contrib.admin.models import LogEntry, ADDITION, DELETION


unicode =str
import xlwt
from xlwt import *


@login_required(redirect_field_name='ret', login_url='/loginsga')
@secure_module
@last_access
@transaction.atomic()
def view(request):
    data = {}
    data['personasesion']=personasesion = request.session['persona']
    lista_carreras = Inscripcion.objects.values_list('carrera__id').filter(status=True, carrera__status=True, carrera__modalidad=3,coordinacion__excluir=True).distinct()
    data['lista_carreras'] = lista_carreras = Carrera.objects.filter(status=True, pk__in=lista_carreras)
    data['periodo'] = periodo = request.session['periodo']
    data['cantidad_graduados_online'] = cantidad_graduados_online = Matricula.objects.values("id").filter(inscripcion__carrera__modalidad=3, nivel__periodo=periodo).distinct().count()
    data['coordinacion'] = coordinacion =  Coordinacion.objects.get(status=True,id=9)
    adduserdata(request, data)
    if request.method == 'POST':
        action = request.POST['action']

        if action == 'add_titulos_profesionales':
            try:
                f = PersonaTituloUniversidadForm(request.POST)
                if f.is_valid():
                    persona = Persona.objects.get(pk=request.POST['idp'])
                    personatitulouniversidad = PersonaTituloUniversidad(persona=persona,
                                                                        codigoregistro=f.cleaned_data['codigoregistro'],
                                                                        fecharegistro=f.cleaned_data['fecharegistro'],
                                                                        fechaacta=f.cleaned_data['fechaacta'],
                                                                        fechainicio=f.cleaned_data['fechainicio'],
                                                                        fecharegresado=f.cleaned_data['fecharegresado'],
                                                                        universidad=f.cleaned_data['universidad'],
                                                                        tipouniversidad=f.cleaned_data['tipouniversidad'],
                                                                        tiponivel=f.cleaned_data['tiponivel'],
                                                                        nombrecarrera=f.cleaned_data['nombrecarrera'])
                    personatitulouniversidad.save(request)
                    log(u'Agrego titulo profesional: %s' % personatitulouniversidad, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                     raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'edit_titulos_profesionales':
            try:
                f = PersonaTituloUniversidadForm(request.POST)
                if f.is_valid():
                    personatitulouniversidad = PersonaTituloUniversidad.objects.get(pk=request.POST['id'])
                    personatitulouniversidad.codigoregistro = f.cleaned_data['codigoregistro']
                    personatitulouniversidad.fecharegistro = f.cleaned_data['fecharegistro']
                    personatitulouniversidad.fechaacta = f.cleaned_data['fechaacta']
                    personatitulouniversidad.fechainicio = f.cleaned_data['fechainicio']
                    personatitulouniversidad.fecharegresado = f.cleaned_data['fecharegresado']
                    personatitulouniversidad.universidad = f.cleaned_data['universidad']
                    personatitulouniversidad.tipouniversidad = f.cleaned_data['tipouniversidad']
                    personatitulouniversidad.tiponivel = f.cleaned_data['tiponivel']
                    personatitulouniversidad.nombrecarrera = f.cleaned_data['nombrecarrera']
                    personatitulouniversidad.save(request)
                    log(u'Edito titulo profesional: %s' % personatitulouniversidad, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                     raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'delete_titulos_profesionales':
            try:
                personatitulouniversidad = PersonaTituloUniversidad.objects.get(pk=request.POST['id'])
                personatitulouniversidad.delete()
                log(u'Elimino titulo profesional: %s' % personatitulouniversidad, request, "del")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'cambiomalla':
            try:
                inscripcion = Inscripcion.objects.get(pk=request.POST['id'])
                f = CambiomallaForm(request.POST)
                if f.is_valid():
                    malla = inscripcion.inscripcionmalla_set.filter(status=True)
                    malla.delete()
                    im = InscripcionMalla(inscripcion=inscripcion,
                                          malla=f.cleaned_data['malla_nueva'])
                    im.save(request)
                    inscripcion.actualizar_creditos()
                    inscripcion.actualizar_nivel()
                    log(u'Modifico malla de inscripcion: %s - %s' % (inscripcion.persona, im.malla), request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'cambionivel':
            try:
                inscripcion = Inscripcion.objects.get(pk=request.POST['id'])
                f = CambionivelmallaForm(request.POST)
                if f.is_valid():
                    nivel = inscripcion.mi_nivel()
                    nivel.nivel = f.cleaned_data['nuevonivel']
                    nivel.save(request)
                    log(u'Modifico nivel de inscripcion: %s - %s' % (inscripcion.persona, nivel.nivel), request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'cambiocategoria':
            try:
                inscripcion = Inscripcion.objects.get(pk=request.POST['id'])
                f = InscripcionTipoInscripcionForm(request.POST)
                if f.is_valid():
                    if inscripcion.tipo_inscripcion():
                        tipo = inscripcion.tipo_inscripcion()
                        tipo.tipoinscripcion = f.cleaned_data['tipoinscripcion']
                        tipo.save(request)
                    else:
                        tipo = InscripcionTipoInscripcion(inscripcion=inscripcion,
                                                          tipoinscripcion=f.cleaned_data['tipoinscripcion'])
                        tipo.save(request)
                    log(u'Modifico categoria de inscripcion: %s' % inscripcion.persona, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'chequeoinscripcionedit':
            try:
                cedula = request.POST['cedula']
                carrera = request.POST['carrera']
                idi = request.POST['id']
                if cedula.__len__() > 0 and carrera.__len__() > 0:
                    if not Inscripcion.objects.filter(persona__cedula=request.POST['cedula'], carrera=Carrera.objects.get(pk=carrera)):
                        return JsonResponse({"result": "ok"})
                    else:
                        ins = Inscripcion.objects.filter(persona__cedula=request.POST['cedula'], carrera=Carrera.objects.get(pk=carrera))[0]
                        if ins.id != idi:
                            return JsonResponse({"result": "bad"})
                        else:
                            return JsonResponse({"result": "ok"})
                else:
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                return JsonResponse({"result": "bad"})

        elif action == 'addrecord':
            try:
                inscripcion = Inscripcion.objects.get(pk=request.POST['id'])
                f = RecordAcademicoForm(request.POST)
                if f.is_valid():
                    if inscripcion.recordacademico_set.filter(asignatura=f.cleaned_data['asignatura']).exists():
                        return JsonResponse({"result": "bad", "mensaje": u"Ya existe la asignatura en el record, modifiquela desde el historico."})
                    record = RecordAcademico(inscripcion=inscripcion,
                                             asignatura=f.cleaned_data['asignatura'],
                                             modulomalla=inscripcion.asignatura_en_modulomalla(f.cleaned_data['asignatura']),
                                             asignaturamalla=inscripcion.asignatura_en_asignaturamalla(f.cleaned_data['asignatura']),
                                             nota=f.cleaned_data['nota'],
                                             asistencia=f.cleaned_data['asistencia'],
                                             fecha=f.cleaned_data['fecha'],
                                             aprobada=f.cleaned_data['aprobada'],
                                             noaplica=f.cleaned_data['noaplica'],
                                             convalidacion=f.cleaned_data['convalidacion'],
                                             pendiente=False,
                                             creditos=f.cleaned_data['creditos'],
                                             horas=f.cleaned_data['horas'],
                                             homologada=f.cleaned_data['homologada'],
                                             valida=f.cleaned_data['valida'],
                                             validapromedio=f.cleaned_data['validapromedio'],
                                             observaciones=f.cleaned_data['observaciones'])
                    record.save(request)
                    record.actualizar()
                    inscripcion.actualizar_nivel()
                    inscripcion.actualiza_matriculas(record.asignatura)
                    log(u'Adiciono record academico: %s - %s' % (record, record.inscripcion.persona), request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos. "})

        elif action == 'addrecordhomologada':
            try:
                inscripcion = Inscripcion.objects.get(pk=request.POST['id'])
                f = RecordAcademicoForm(request.POST)
                if f.is_valid():
                    if not inscripcion.recordacademico_set.filter(asignatura=f.cleaned_data['asignatura']).exists():
                        record = RecordAcademico(inscripcion=inscripcion,
                                                 asignatura=f.cleaned_data['asignatura'],
                                                 modulomalla=inscripcion.asignatura_en_modulomalla(f.cleaned_data['asignatura']),
                                                 asignaturamalla=inscripcion.asignatura_en_asignaturamalla(f.cleaned_data['asignatura']),
                                                 nota=f.cleaned_data['nota'],
                                                 fecha=f.cleaned_data['fecha'],
                                                 aprobada=True,
                                                 noaplica=f.cleaned_data['noaplica'],
                                                 convalidacion=f.cleaned_data['convalidacion'],
                                                 pendiente=False,
                                                 valida=f.cleaned_data['valida'],
                                                 validapromedio=f.cleaned_data['validapromedio'],
                                                 creditos=f.cleaned_data['creditos'],
                                                 horas=f.cleaned_data['horas'],
                                                 homologada=f.cleaned_data['homologada'],
                                                 observaciones=f.cleaned_data['observaciones'])
                        record.save(request)
                    else:
                        recordexistente = inscripcion.recordacademico_set.filter(asignatura=f.cleaned_data['asignatura'])[0]
                        if recordexistente.historicorecordacademico_set.filter(fecha=f.cleaned_data['fecha']).exists():
                            return JsonResponse({"result": "bad", "mensaje": u"Registro existente en esa fecha"})
                        record = HistoricoRecordAcademico(recordacademico=recordexistente,
                                                          inscripcion=inscripcion,
                                                          asignatura=f.cleaned_data['asignatura'],
                                                          nota=f.cleaned_data['nota'],
                                                          fecha=f.cleaned_data['fecha'],
                                                          aprobada=True,
                                                          noaplica=f.cleaned_data['noaplica'],
                                                          convalidacion=f.cleaned_data['convalidacion'],
                                                          pendiente=False,
                                                          valida=f.cleaned_data['valida'],
                                                          validapromedio=f.cleaned_data['validapromedio'],
                                                          creditos=f.cleaned_data['creditos'],
                                                          horas=f.cleaned_data['horas'],
                                                          homologada=f.cleaned_data['homologada'],
                                                          observaciones=f.cleaned_data['observaciones'])
                        record.save(request)
                    record.actualizar()
                    inscripcion.actualizar_nivel()
                    inscripcion.actualiza_matriculas(f.cleaned_data['asignatura'])
                    log(u'Adiciono record academico: %s - %s' % (record, record.inscripcion.persona), request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos. "})

        elif action == 'addhistorico':
            try:
                record = RecordAcademico.objects.get(pk=request.POST['id'])
                f = HistoricoRecordAcademicoForm(request.POST)
                if f.is_valid():
                    if record.historicorecordacademico_set.filter(fecha=f.cleaned_data['fecha']).exists():
                        return JsonResponse({"result": "bad", "mensaje": u"Registro existente en esa fecha"})
                    historico = HistoricoRecordAcademico(recordacademico=record,
                                                         inscripcion=record.inscripcion,
                                                         asignatura=f.cleaned_data['asignatura'],
                                                         modulomalla=record.inscripcion.asignatura_en_modulomalla(f.cleaned_data['asignatura']),
                                                         asignaturamalla=record.inscripcion.asignatura_en_asignaturamalla(f.cleaned_data['asignatura']),
                                                         nota=f.cleaned_data['nota'],
                                                         asistencia=f.cleaned_data['asistencia'],
                                                         fecha=f.cleaned_data['fecha'],
                                                         aprobada=f.cleaned_data['aprobada'],
                                                         convalidacion=f.cleaned_data['convalidacion'],
                                                         noaplica=f.cleaned_data['noaplica'],
                                                         pendiente=False,
                                                         creditos=f.cleaned_data['creditos'],
                                                         horas=f.cleaned_data['horas'],
                                                         homologada=f.cleaned_data['homologada'],
                                                         valida=f.cleaned_data['valida'],
                                                         validapromedio=f.cleaned_data['validapromedio'],
                                                         observaciones=f.cleaned_data['observaciones'])
                    historico.save(request)
                    historico.actualizar()
                    record.inscripcion.actualizar_nivel()
                    record.inscripcion.actualiza_matriculas(record.asignatura)
                    log(u'Adiciono historico y registro: %s - %s' % (historico, historico.inscripcion.persona), request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'edithistorico':
            try:
                historico = HistoricoRecordAcademico.objects.get(pk=request.POST['id'])
                f = HistoricoRecordAcademicoForm(request.POST)
                if f.is_valid():
                    if HistoricoRecordAcademico.objects.filter(inscripcion=historico.inscripcion, asignatura=historico.asignatura, fecha=f.cleaned_data['fecha']).exclude(id=historico.id).exists():
                        return JsonResponse({"result": "bad", "mensaje": u"Registro existente en esa fecha."})
                    historico.fecha = f.cleaned_data['fecha']
                    if not historico.tiene_acta_nivel():
                        historico.nota = f.cleaned_data['nota']
                        historico.asistencia = f.cleaned_data['asistencia']
                    historico.aprobada = f.cleaned_data['aprobada']
                    historico.noaplica = f.cleaned_data['noaplica']
                    historico.creditos = f.cleaned_data['creditos']
                    historico.horas = f.cleaned_data['horas']
                    historico.homologada = f.cleaned_data['homologada']
                    historico.convalidacion = f.cleaned_data['convalidacion']
                    historico.observaciones = f.cleaned_data['observaciones']
                    historico.valida = f.cleaned_data['valida']
                    historico.validapromedio = f.cleaned_data['validapromedio']
                    historico.save(request)
                    historico.actualizar()
                    historico.inscripcion.actualizar_nivel()
                    historico.inscripcion.actualiza_matriculas(historico.asignatura)
                    log(u'Modifico historico de record academico: %s - %s' % (historico, historico.inscripcion.persona), request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'delrecord':
            try:
                record = RecordAcademico.objects.get(pk=request.POST['id'])
                asignatura = record.asignatura
                inscripcion = record.inscripcion
                homologacion = record.homologacioninscripcion_set.all()
                homologacion.delete()
                convalidacion = record.convalidacioninscripcion_set.all()
                convalidacion.delete()
                historico = record.historicorecordacademico_set.all()
                historico.delete()
                log(u'Elimino registro academico: %s - %s' % (record, record.inscripcion.persona), request, "del")
                record.delete()
                inscripcion.actualizar_nivel()
                inscripcion.actualiza_matriculas(asignatura)
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'delhistorico':
            try:
                historico = HistoricoRecordAcademico.objects.get(pk=request.POST['id'])
                asignatura = historico.asignatura
                inscripcion = historico.inscripcion
                if historico.convalidacion:
                    convalidacion = historico.convalidacioninscripcion_set.all()
                    convalidacion.delete()
                if historico.homologada:
                    homologacion = historico.homologacioninscripcion_set.all()
                    homologacion.delete()
                log(u'Elimino historico de registro academico: %s - %s' % (historico, historico.inscripcion.persona), request, "del")
                historico.delete()
                if HistoricoRecordAcademico.objects.filter(asignatura=asignatura, inscripcion=inscripcion).exists():
                    historico = HistoricoRecordAcademico.objects.filter(asignatura=asignatura, inscripcion=inscripcion)[0]
                    historico.actualizar()
                else:
                    record = RecordAcademico.objects.filter(asignatura=asignatura, inscripcion=inscripcion)
                    record.delete()
                inscripcion.actualizar_nivel()
                inscripcion.actualiza_matriculas(asignatura)
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'convalidar':
            try:
                convalidacion = ConvalidacionInscripcion.objects.get(pk=request.POST['id'])
                f = ConvalidacionInscripcionForm(request.POST, request.FILES)
                if f.is_valid():
                    convalidacion.centro = f.cleaned_data['centro']
                    convalidacion.carrera = f.cleaned_data['carrera']
                    convalidacion.asignatura = f.cleaned_data['asignatura']
                    convalidacion.anno = f.cleaned_data['anno']
                    convalidacion.nota_ant = f.cleaned_data['nota_ant']
                    convalidacion.nota_act = f.cleaned_data['nota_act']
                    convalidacion.observaciones = f.cleaned_data['observaciones']
                    convalidacion.creditos = f.cleaned_data['creditos']
                    convalidacion.save(request)
                    if 'archivo' in request.FILES:
                        archivo = request.FILES['archivo']
                        archivo._name = generar_nombre("archivo_", archivo._name)
                        convalidacion.archivo = archivo
                        convalidacion.save(request)
                    log(u'Adiciono convalidacion: %s - %s' % (convalidacion, convalidacion.record.inscripcion.persona), request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'homologar':
            try:
                homologacion = HomologacionInscripcion.objects.get(pk=request.POST['id'])
                f = HomologacionInscripcionForm(request.POST, request.FILES)
                if f.is_valid():
                    homologacion.carrera = f.cleaned_data['carrera']
                    homologacion.modalidad = f.cleaned_data['modalidad']
                    homologacion.asignatura = f.cleaned_data['asignatura']
                    homologacion.fecha = f.cleaned_data['fecha']
                    homologacion.nota_ant = f.cleaned_data['nota_ant']
                    homologacion.observaciones = f.cleaned_data['observaciones']
                    homologacion.creditos = f.cleaned_data['creditos']
                    homologacion.save(request)
                    if 'archivo' in request.FILES:
                        archivo = request.FILES['archivo']
                        archivo._name = generar_nombre("archivo_", archivo._name)
                        homologacion.archivo = archivo
                        homologacion.save(request)
                    log(u'Adiciono homologacion: %s - %s' % (homologacion, homologacion.record.inscripcion.persona), request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'eliminarmasivo':
            try:
                datos = json.loads(request.POST['lista'])
                if not datos:
                    return JsonResponse({"result": "bad", "mensaje": u"Debe seleccionar datos"})
                for elemento in datos:
                    record = RecordAcademico.objects.get(pk=int(elemento['id']))
                    asignatura = record.asignatura
                    inscripcion = record.inscripcion
                    homologacion = record.homologacioninscripcion_set.all()
                    homologacion.delete()
                    convalidacion = record.convalidacioninscripcion_set.all()
                    convalidacion.delete()
                    historico = record.historicorecordacademico_set.all()
                    historico.delete()
                    record.delete()
                    inscripcion.actualizar_nivel()
                    inscripcion.actualiza_matriculas(asignatura)
                log(u'Elimino materias de record: %s - %s' % (personasesion, record.inscripcion.persona), request, "add")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == "enviomailindividualonline":
            try:
                inscripcion = Inscripcion.objects.get(pk=int(encrypt(request.POST['idinscripcionid'])))
                if request.POST['actualiza'] == '1':
                    inscripcion.persona.email = request.POST['correo']
                    inscripcion.persona.save(request)
                inscripcion.envio_correo_admision(20, request)
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                pass

        elif action == "enviomailindividualonlinenota":
            try:
                inscripcion = Inscripcion.objects.get(pk=int(encrypt(request.POST['idinscripcionid'])))
                materia=None
                if 'idmateria' in request.POST:
                    idmateria = int(request.POST['idmateria'])
                    if idmateria > 0:
                        materia = Materia.objects.get(pk=idmateria)
                if request.POST['actualiza'] == '1':
                    inscripcion.persona.email = request.POST['correo']
                    inscripcion.persona.save(request)
                inscripcion.envio_nota_correo_admision(20,materia,inscripcion, request)
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                pass

        elif action == 'imprimirnotaexamen':
            try:
                data['title'] = u'NOTA FINAL DEL EXÁMEN'
                inscripcion = Inscripcion.objects.get(pk=int(encrypt(request.POST['idinscripcionid'])))
                materia = None
                if 'idmateria' in request.POST:
                    idmateria = int(request.POST['idmateria'])
                    if idmateria > 0:
                        materia = Materia.objects.get(pk=idmateria)
                # materia = Materia.objects.get(pk=14777)
                return conviert_html_to_pdf('inscripciones_admision/imprimircertificadonota.html',
                                            {'pagesize': 'A4',
                                             'data': data,
                                             'inscripcion': inscripcion,
                                             'materia': materia,
                                             })
            except Exception as ex:
                pass

        elif action == 'imprimir_horario_estudiante':
            try:
                data['title'] = u'HORARIO DE EXAMEN'
                if 'sede' in request.POST:
                    sede=request.POST['sede']
                if 'dia' in request.POST:
                    dia=request.POST['dia']
                if 'hora_inicio' in request.POST:
                    hora_inicio=request.POST['hora_inicio']
                if 'hora_fin' in request.POST:
                    hora_fin=request.POST['hora_fin']
                if 'bloque' in request.POST:
                    bloque=request.POST['bloque']
                if 'lab' in request.POST:
                    lab=request.POST['lab']

                sql = "select distinct est.cedula,est.apel,est.nom from horario h inner JOIN ficha_estudiante est on h.cedula=est.cedula INNER JOIN carrera car on est.id_carrera=car.id_carrera " \
                      "INNER JOIN asignatura asig on est.id_asigantura=asig.id_asigantura " \
                      "WHERE h.sede='"+sede+"' AND h.dia='"+dia+"'  and (h.horario = '"+hora_inicio+"' or h.horario='"+hora_fin+"') and h.laboratorio='"+lab+"' and h.bloque='"+bloque+"' ORDER BY est.apel,est.nom"

                listado_alumnos = querymysqlconsulta(sql,True)
                return conviert_html_to_pdf('inscripciones_admision/horario_estudiantes.html',
                                            {'pagesize': 'A4',
                                             'data': data,
                                             'listado_alumnos': listado_alumnos,
                                             'dia': dia,
                                             'hora_inicio': hora_inicio,
                                             'hora_fin': hora_fin,
                                             'bloque': bloque,
                                             'laboratorio': lab,
                                             'sede': sede,
                                             })
            except Exception as ex:
                pass

        elif action == 'cargarfoto':
            try:
                inscripcion = Inscripcion.objects.get(pk=request.POST['id'])
                form = CargarFotoForm(inscripcion, request.FILES)
                if form.is_valid():
                    persona = inscripcion.persona
                    newfile = request.FILES['foto']
                    newfile._name = generar_nombre("foto_", newfile._name)
                    foto = persona.foto()
                    if foto:
                        foto.foto = newfile
                    else:
                        foto = FotoPersona(persona=persona,
                                           foto=newfile)
                    foto.save(request)
                    make_thumb_picture(persona)
                    if GENERAR_TUMBAIL:
                        make_thumb_fotopersona(persona)
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos. "})

        elif action == 'cambiogrupo':
            try:
                inscripcion = Inscripcion.objects.get(pk=request.POST['id'])
                if inscripcion.matriculado():
                    return JsonResponse({"result": "bad", "mensaje": u"Existe una matricula activa."})
                f = CambioGrupoForm(request.POST)
                if f.is_valid():
                    inscripcion.inscripcion_grupo(f.cleaned_data['grupo'])
                    log(u'Cambio de grupo: %s - %s' % (inscripcion.persona, f.cleaned_data['grupo']), request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        # elif action == 'add':
        #     try:
        #         f = InscripcionForm(request.POST)
        #         isPPL = False
        #         archivoppl = None
        #         if 'ppl' in request.POST:
        #             isPPL = True
        #             f.fields['fechaingresoppl'].required = True
        #             if 'archivoppl' in request.FILES:
        #                 arch = request.FILES['archivoppl']
        #                 extension = arch._name.split('.')
        #                 tam = len(extension)
        #                 exte = extension[tam - 1]
        #                 if arch.size > 4194304:
        #                     raise NameError(u"Tamaño del archivo es mayor a 4 Mb")
        #                 if not exte.lower() == 'pdf':
        #                     raise NameError(u"Solo se permiten archivos .pdf")
        #                 archivoppl = request.FILES['archivoppl']
        #                 archivoppl._name = generar_nombre("archivoppl_", archivoppl._name)
        #         if not f.is_valid():
        #             raise NameError('Formulario incorrecto')
        #         if f.cleaned_data['cedula'] and Persona.objects.filter(cedula=f.cleaned_data['cedula']).exists():
        #             if Inscripcion.objects.filter(persona__cedula=f.cleaned_data['cedula'], carrera=f.cleaned_data['carrera']).exists():
        #                 return JsonResponse({"result": "bad", "mensaje": u"El numero de cedula o pasaporte ya esta registrado en la carrera seleccionada."})
        #         if f.cleaned_data['pasaporte'] and Persona.objects.filter(pasaporte=f.cleaned_data['pasaporte']).exists():
        #             return JsonResponse({"result": "bad", "mensaje": u"El numero de cedula o pasaporte ya esta registrado."})
        #         if not f.cleaned_data['cedula'] and not f.cleaned_data['pasaporte']:
        #             return JsonResponse({"result": "bad", "mensaje": u"Debe especificar un numero de identificación."})
        #
        #         # if Persona.objects.filter(cedula=f.cleaned_data['cedula']).exists():
        #         if Persona.objects.filter(Q(cedula=f.cleaned_data['cedula']) & Q(pasaporte=f.cleaned_data['pasaporte'])).exists():
        #
        #             if f.cleaned_data['cedula']:
        #                 persona = Persona.objects.get(cedula=f.cleaned_data['cedula'])
        #             else:
        #                 persona = Persona.objects.get(pasaporte=f.cleaned_data['pasaporte'])
        #             persona.nombres = f.cleaned_data['nombres']
        #             persona.apellido1 = f.cleaned_data['apellido1']
        #             persona.apellido2 = f.cleaned_data['apellido2']
        #             persona.cedula = f.cleaned_data['cedula']
        #             persona.pasaporte = f.cleaned_data['pasaporte']
        #             persona.nacimiento = f.cleaned_data['nacimiento']
        #             persona.sexo = f.cleaned_data['sexo']
        #             persona.paisnacimiento = f.cleaned_data['paisnacimiento']
        #             persona.provincianacimiento = f.cleaned_data['provincianacimiento']
        #             persona.cantonnacimiento = f.cleaned_data['cantonnacimiento']
        #             persona.parroquianacimiento = f.cleaned_data['parroquianacimiento']
        #             persona.nacionalidad = f.cleaned_data['nacionalidad']
        #             persona.pais = f.cleaned_data['pais']
        #             persona.provincia = f.cleaned_data['provincia']
        #             persona.canton = f.cleaned_data['canton']
        #             persona.parroquia = f.cleaned_data['parroquia']
        #             persona.sector = f.cleaned_data['sector']
        #             persona.direccion = f.cleaned_data['direccion']
        #             persona.direccion2 = f.cleaned_data['direccion2']
        #             persona.num_direccion = f.cleaned_data['num_direccion']
        #             persona.telefono = f.cleaned_data['telefono']
        #             persona.telefono_conv = f.cleaned_data['telefono_conv']
        #             persona.email = f.cleaned_data['email']
        #             persona.sangre = f.cleaned_data['sangre']
        #             persona.lgtbi = f.cleaned_data['lgtbi']
        #             persona.save(request)
        #             if not persona.usuario:
        #                 username = calculate_username(persona)
        #                 generar_usuario(persona, username, ALUMNOS_GROUP_ID)
        #             else:
        #                 username = persona.usuario.username
        #         else:
        #             persona = Persona(nombres=f.cleaned_data['nombres'],
        #                               apellido1=f.cleaned_data['apellido1'],
        #                               apellido2=f.cleaned_data['apellido2'],
        #                               cedula=f.cleaned_data['cedula'],
        #                               pasaporte=f.cleaned_data['pasaporte'],
        #                               nacimiento=f.cleaned_data['nacimiento'],
        #                               sexo=f.cleaned_data['sexo'],
        #                               paisnacimiento=f.cleaned_data['paisnacimiento'],
        #                               provincianacimiento=f.cleaned_data['provincianacimiento'],
        #                               cantonnacimiento=f.cleaned_data['cantonnacimiento'],
        #                               parroquianacimiento=f.cleaned_data['parroquianacimiento'],
        #                               nacionalidad=f.cleaned_data['nacionalidad'],
        #                               pais=f.cleaned_data['pais'],
        #                               provincia=f.cleaned_data['provincia'],
        #                               canton=f.cleaned_data['canton'],
        #                               parroquia=f.cleaned_data['parroquia'],
        #                               sector=f.cleaned_data['sector'],
        #                               direccion=f.cleaned_data['direccion'],
        #                               direccion2=f.cleaned_data['direccion2'],
        #                               num_direccion=f.cleaned_data['num_direccion'],
        #                               telefono=f.cleaned_data['telefono'],
        #                               telefono_conv=f.cleaned_data['telefono_conv'],
        #                               email=f.cleaned_data['email'],
        #                               sangre=f.cleaned_data['sangre'],
        #                               lgtbi=f.cleaned_data['lgtbi'])
        #             persona.save(request)
        #             username = calculate_username(persona)
        #             generar_usuario(persona, username, ALUMNOS_GROUP_ID)
        #         if EMAIL_INSTITUCIONAL_AUTOMATICO:
        #             persona.emailinst = username + '@' + EMAIL_DOMAIN
        #             persona.save(request)
        #         if UTILIZA_GRUPOS_ALUMNOS:
        #             grupo = f.cleaned_data['grupo']
        #             carrera = grupo.carrera
        #             sesion = grupo.sesion
        #             modalidad = grupo.modalidad
        #             sede = grupo.sede
        #         else:
        #             carrera = f.cleaned_data['carrera']
        #             sesion = f.cleaned_data['sesion']
        #             modalidad = f.cleaned_data['modalidad']
        #             sede = f.cleaned_data['sede']
        #         inscripcion = Inscripcion(persona=persona,
        #                                   fecha=f.cleaned_data['fecha'],
        #                                   especialidad=f.cleaned_data['especialidad'],
        #                                   identificador=f.cleaned_data['identificador'],
        #                                   centroinformacion=f.cleaned_data['centroinformacion'],
        #                                   carrera=carrera,
        #                                   modalidad=modalidad,
        #                                   sesion=sesion,
        #                                   sede=sede)
        #         if f.cleaned_data['unidadeducativa']:
        #             inscripcion.unidadeducativa_id=f.cleaned_data['unidadeducativa']
        #         inscripcion.save(request)
        #         if isPPL:
        #             if HistorialPersonaPPL.objects.filter(fechaingreso=f.cleaned_data['fechaingresoppl'], persona=persona, inscripcion=inscripcion).exists():
        #                 historialppl = HistorialPersonaPPL.objects.filter(fechaingreso=f.cleaned_data['fechaingresoppl'], persona=persona, inscripcion=inscripcion).first()
        #                 historialppl.observacion = f.cleaned_data['observacionppl'] if f.cleaned_data['observacionppl'] else historialppl.observacion
        #                 historialppl.archivo = archivoppl if archivoppl else historialppl.archivo
        #                 historialppl.centrorehabilitacion = f.cleaned_data['centrorehabilitacion'] if f.cleaned_data['centrorehabilitacion'] else historialppl.centrorehabilitacion
        #                 historialppl.lidereducativo = f.cleaned_data['lidereducativo'] if f.cleaned_data['lidereducativo'] else historialppl.lidereducativo
        #                 historialppl.correolidereducativo = f.cleaned_data['correolidereducativo'] if f.cleaned_data['correolidereducativo'] else historialppl.correolidereducativo
        #                 historialppl.telefonolidereducativo = f.cleaned_data['telefonolidereducativo'] if f.cleaned_data['telefonolidereducativo'] else historialppl.telefonolidereducativo
        #             else:
        #                 historialppl = HistorialPersonaPPL(persona=persona,
        #                                                    inscripcion=inscripcion,
        #                                                    observacion=f.cleaned_data['observacionppl'] if f.cleaned_data['observacionppl'] else None,
        #                                                    archivo=archivoppl,
        #                                                    fechaingreso=f.cleaned_data['fechaingresoppl'],
        #                                                    centrorehabilitacion=f.cleaned_data['centrorehabilitacion'] if f.cleaned_data['centrorehabilitacion'] else None,
        #                                                    lidereducativo=f.cleaned_data['lidereducativo'] if f.cleaned_data['lidereducativo'] else None,
        #                                                    correolidereducativo=f.cleaned_data['correolidereducativo'] if f.cleaned_data['correolidereducativo'] else None,
        #                                                    telefonolidereducativo=f.cleaned_data['telefonolidereducativo'] if f.cleaned_data['telefonolidereducativo'] else None,
        #                                                    )
        #             historialppl.save(request)
        #         if not inscripcion.coordinacion:
        #             transaction.set_rollback(True)
        #             return JsonResponse({"result": "bad", "mensaje": u"No existe una coordinación para esta carrera en esta sede."})
        #         # DOCUMENTOS DE INSCRIPCION
        #         documentos = DocumentosDeInscripcion(inscripcion=inscripcion,
        #                                              titulo=f.cleaned_data['titulo'],
        #                                              acta=f.cleaned_data['acta'],
        #                                              cedula=f.cleaned_data['cedula2'],
        #                                              votacion=f.cleaned_data['votacion'],
        #                                              actaconv=f.cleaned_data['actaconv'],
        #                                              partida_nac=f.cleaned_data['partida_nac'],
        #                                              pre=f.cleaned_data['prenivelacion'],
        #                                              observaciones_pre=f.cleaned_data['observacionespre'],
        #                                              fotos=f.cleaned_data['fotos'])
        #         documentos.save(request)
        #         # DOCUMENTOS CONDUCCION
        #         inscripciontesdrive = InscripcionTesDrive(inscripcion=inscripcion,
        #                                                   licencia=f.cleaned_data['licencia'],
        #                                                   record=f.cleaned_data['record'],
        #                                                   certificado_tipo_sangre=f.cleaned_data['certificado_tipo_sangre'],
        #                                                   prueba_psicosensometrica=f.cleaned_data['prueba_psicosensometrica'],
        #                                                   certificado_estudios=f.cleaned_data['certificado_estudios'])
        #         inscripciontesdrive.save(request)
        #         # SEGUIMIENTO LABORAL
        #         if f.cleaned_data['trabaja']:
        #             trabajo = SeguimientoEstudiante(persona=persona,
        #                                             empresa=f.cleaned_data['empresa'],
        #                                             ocupacion=f.cleaned_data['ocupacion'],
        #                                             responsabilidades='',
        #                                             personascargo=0,
        #                                             ejerce=False,
        #                                             fecha=f.cleaned_data['fecha_ingreso'],
        #                                             telefono=f.cleaned_data['telefono_trabajo'])
        #             trabajo.save(request)
        #         # REGISTRO TIPO DE INSCRIPCION
        #         if USA_TIPOS_INSCRIPCIONES:
        #             inscripciontipoinscripcion = InscripcionTipoInscripcion(inscripcion=inscripcion,
        #                                                                     tipoinscripcion_id=TIPO_INSCRIPCION_INICIAL)
        #             inscripciontipoinscripcion.save(request)
        #         # PREGUNTAS
        #         preguntasinscripcion = inscripcion.preguntas_inscripcion()
        #         if f.cleaned_data['comoseinformo']:
        #             preguntasinscripcion.comoseinformo = f.cleaned_data['comoseinformo']
        #         if f.cleaned_data['razonesmotivaron']:
        #             preguntasinscripcion.razonesmotivaron = f.cleaned_data['razonesmotivaron']
        #         if f.cleaned_data['comoseinformootras']:
        #             preguntasinscripcion.comoseinformootras = f.cleaned_data['comoseinformootras']
        #         preguntasinscripcion.save(request)
        #         # PREINSCRITO
        #         if 'preinscrito_id' in request.POST:
        #             preinscrito = PreInscrito.objects.get(pk=request.POST['preinscrito_id'])
        #             preinscrito.inscripcion = inscripcion
        #             preinscrito.save(request)
        #         # PERFIL DE USUARIO
        #         persona.crear_perfil(inscripcion=inscripcion)
        #         perfil=persona.mi_perfil()
        #         perfil.raza=(f.cleaned_data['raza'])
        #         perfil.nacionalidadindigena = (f.cleaned_data['nacionalidadindigena'])
        #         perfil.save(request)
        #         if UTILIZA_GRUPOS_ALUMNOS:
        #             inscripcion.inscripcion_grupo(f.cleaned_data['grupo'])
        #         inscripcion.persona.mi_perfil()
        #         inscripcion.malla_inscripcion()
        #         inscripcion.actualizar_nivel()
        #         fichasocioeconomica = ficha_socioeconomica(inscripcion.persona)
        #         log(u'Adiciono inscripcion: %s' % inscripcion, request, "add")
        #         #send_html_mail("Crear nueva cuenta de correo ", "emails/nuevacuentacorreo.html", {'sistema': request.session['nombresistema'], 'persona': persona, 't': miinstitucion(), 'tipo_usuario': 'ESTUDIANTE', 'inscripcion': inscripcion}, lista_correo([variable_valor('ADMINISTRADOR_CORREO_GROUP_ID')]), [], cuenta=CUENTAS_CORREOS[4][1])
        #         persona.creacion_persona(request.session['nombresistema'],personasesion)
        #         return JsonResponse({"result": "ok", "id": inscripcion.id})
        #     except Exception as ex:
        #         transaction.set_rollback(True)
        #         return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos. %s" % ex.__str__()})

        # elif action == 'adicionarotracarrera':
        #     try:
        #         f = NuevaInscripcionForm(request.POST)
        #         if f.is_valid() and not Inscripcion.objects.filter(id=int(request.POST['id']), carrera=f.cleaned_data['carrera']).exists():
        #             inscripcion = Inscripcion.objects.get(pk=int(request.POST['id']))
        #             carrera = f.cleaned_data['carrera']
        #             sesion = f.cleaned_data['sesion']
        #             modalidad = f.cleaned_data['modalidad']
        #             sede = f.cleaned_data['sede']
        #             if Inscripcion.objects.filter(persona=inscripcion.persona, carrera=carrera).exists():
        #                 return JsonResponse({"result": "bad", "mensaje": u"Ya se encuentra registrado en esa carrera."})
        #             nuevainscripcion = Inscripcion(persona=inscripcion.persona,
        #                                            fecha=f.cleaned_data['fecha'],
        #                                            colegio=inscripcion.colegio,
        #                                            especialidad=inscripcion.especialidad,
        #                                            carrera=carrera,
        #                                            modalidad=modalidad,
        #                                            sesion=sesion,
        #                                            sede=sede)
        #             nuevainscripcion.save(request)
        #             if f.cleaned_data['copiarecord']:
        #                 for record in inscripcion.recordacademico():
        #                     nuevorecord = RecordAcademico(inscripcion=nuevainscripcion,
        #                                                   asignatura=record.asignatura,
        #                                                   nota=record.nota,
        #                                                   asistencia=record.asistencia,
        #                                                   fecha=record.fecha,
        #                                                   noaplica=record.noaplica,
        #                                                   convalidacion=False,
        #                                                   homologada=True,
        #                                                   aprobada=record.aprobada,
        #                                                   pendiente=record.pendiente,
        #                                                   creditos=record.creditos,
        #                                                   horas=record.horas,
        #                                                   valida=record.valida,
        #                                                   observaciones=record.observaciones)
        #                     nuevorecord.save(request)
        #                     nuevorecord.actualizar()
        #             if inscripcion.tipo_inscripcion():
        #                 tipoinscripcion = inscripcion.tipo_inscripcion().tipoinscripcion
        #                 nuevotipo = InscripcionTipoInscripcion(inscripcion=nuevainscripcion,
        #                                                        tipoinscripcion=tipoinscripcion)
        #                 nuevotipo.save(request)
        #             hoy = datetime.now().date()
        #             inscripciontesdrive = InscripcionTesDrive(inscripcion=nuevainscripcion,
        #                                                       licencia=f.cleaned_data['licencia'],
        #                                                       record=f.cleaned_data['record'],
        #                                                       certificado_tipo_sangre=f.cleaned_data['certificado_tipo_sangre'],
        #                                                       prueba_psicosensometrica=f.cleaned_data['prueba_psicosensometrica'],
        #                                                       certificado_estudios=f.cleaned_data['certificado_estudios'])
        #             inscripciontesdrive.save(request)
        #             nuevainscripcion.malla_inscripcion()
        #             log(u'Adiciono inscripcion desde otra carrera: %s' % nuevainscripcion, request, "add")
        #             documentos = DocumentosDeInscripcion(inscripcion=nuevainscripcion,
        #                                                  titulo=f.cleaned_data['titulo'],
        #                                                  acta=f.cleaned_data['acta'],
        #                                                  cedula=f.cleaned_data['cedula2'],
        #                                                  votacion=f.cleaned_data['votacion'],
        #                                                  actaconv=f.cleaned_data['actaconv'],
        #                                                  partida_nac=f.cleaned_data['partida_nac'],
        #                                                  pre=f.cleaned_data['prenivelacion'],
        #                                                  observaciones_pre=f.cleaned_data['observacionespre'],
        #                                                  fotos=f.cleaned_data['fotos'])
        #             documentos.save(request)
        #             inscripcion.persona.crear_perfil(inscripcion=nuevainscripcion)
        #             return JsonResponse({"result": "ok", "id": nuevainscripcion.id})
        #         else:
        #             raise NameError('Error')
        #     except Exception as ex:
        #         transaction.set_rollback(True)
        #         return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'edit':
            try:
                f = InscripcionForm(request.POST)
                if f.is_valid():
                    inscripcion = Inscripcion.objects.get(pk=request.POST['id'])
                    persona = inscripcion.persona
                    perfil = persona.mi_perfil()
                    persona.nombres = f.cleaned_data['nombres']
                    persona.apellido1 = f.cleaned_data['apellido1']
                    persona.apellido2 = f.cleaned_data['apellido2']
                    persona.nacimiento = f.cleaned_data['nacimiento']
                    persona.paisnacimiento = f.cleaned_data['paisnacimiento']
                    persona.provincianacimiento = f.cleaned_data['provincianacimiento']
                    persona.cantonnacimiento = f.cleaned_data['cantonnacimiento']
                    persona.parroquianacimiento = f.cleaned_data['parroquianacimiento']
                    persona.nacionalidad = f.cleaned_data['nacionalidad']
                    persona.pais = f.cleaned_data['pais']
                    persona.provincia = f.cleaned_data['provincia']
                    persona.canton = f.cleaned_data['canton']
                    persona.parroquia = f.cleaned_data['parroquia']
                    persona.sexo = f.cleaned_data['sexo']
                    persona.provincia = f.cleaned_data['provincia']
                    persona.canton = f.cleaned_data['canton']
                    persona.parroquia = f.cleaned_data['parroquia']
                    persona.sector = f.cleaned_data['sector']
                    persona.direccion = f.cleaned_data['direccion']
                    persona.direccion2 = f.cleaned_data['direccion2']
                    persona.num_direccion = f.cleaned_data['num_direccion']
                    persona.telefono = f.cleaned_data['telefono']
                    persona.telefono_conv = f.cleaned_data['telefono_conv']
                    persona.email = f.cleaned_data['email']
                    persona.emailinst = f.cleaned_data['emailinst']
                    persona.sangre = f.cleaned_data['sangre']
                    persona.lgtbi = f.cleaned_data['lgtbi']
                    persona.save(request)
                    #DATOS UBE
                    perfil.raza = f.cleaned_data['raza']
                    perfil.nacionalidadindigena = f.cleaned_data['nacionalidadindigena']
                    perfil.save(request)
                    # DATOS DE LA INSCRIPCION
                    inscripcion.fecha = f.cleaned_data['fecha']
                    if not inscripcion.matriculado() and not inscripcion.graduado() and not inscripcion.egresado():
                        if UTILIZA_GRUPOS_ALUMNOS:
                            grupo = f.cleaned_data['grupo']
                            inscripcion.inscripcion_grupo(grupo)
                        else:
                            inscripcion.modalidad = f.cleaned_data['modalidad']
                            inscripcion.sede = f.cleaned_data['sede']
                    inscripcion.centroinformacion = f.cleaned_data['centroinformacion']
                    inscripcion.sesion = f.cleaned_data['sesion']
                    if f.cleaned_data['unidadeducativa']:
                        if int(f.cleaned_data['unidadeducativa'])>0:
                            inscripcion.unidadeducativa_id=f.cleaned_data['unidadeducativa']
                    inscripcion.especialidad_id = f.cleaned_data['especialidad']
                    inscripcion.identificador = f.cleaned_data['identificador']
                    inscripcion.save(request)
                    # ACTUALIZA LAS PREGUNTAS DE INSCRIPCION
                    preguntas = inscripcion.preguntas_inscripcion()
                    preguntas.comoseinformo = f.cleaned_data['comoseinformo']
                    preguntas.razonesmotivaron = f.cleaned_data['razonesmotivaron']
                    preguntas.comoseinformootras = f.cleaned_data['comoseinformootras']
                    preguntas.save(request)
                    # DOCUMENTOS
                    documentos = inscripcion.documentos_entregados()
                    documentos.acta = f.cleaned_data['acta']
                    documentos.cedula = f.cleaned_data['cedula2']
                    documentos.fotos = f.cleaned_data['fotos']
                    documentos.titulo = f.cleaned_data['titulo']
                    documentos.votacion = f.cleaned_data['votacion']
                    documentos.actaconv = f.cleaned_data['actaconv']
                    documentos.partida_nac = f.cleaned_data['partida_nac']
                    # PRENIVELACION
                    documentos.pre = f.cleaned_data['prenivelacion']
                    documentos.observaciones_pre = f.cleaned_data['observacionespre']
                    documentos.save(request)
                    # OTROS DOCUMENTOS
                    tesdrive = inscripcion.documentos_tesdrive()
                    tesdrive.licencia = f.cleaned_data['licencia']
                    tesdrive.record = f.cleaned_data['record']
                    tesdrive.certificado_tipo_sangre = f.cleaned_data['certificado_tipo_sangre']
                    tesdrive.prueba_psicosensometrica = f.cleaned_data['prueba_psicosensometrica']
                    tesdrive.certificado_estudios = f.cleaned_data['certificado_estudios']
                    tesdrive.save(request)
                    # OTRAS ACCIONES
                    inscripcion.actualizar_nivel()
                    log(u'Modifico de inscripcion: %s' % inscripcion, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'novalidar':
            try:
                record = RecordAcademico.objects.get(pk=request.POST['id'])
                f = ConsiderarForm(request.POST)
                if f.is_valid():
                    motivo = f.cleaned_data['motivo']
                    record.valida = False
                    record.save(request)
                    historico = record.mi_historico()
                    historico.valida = record.valida
                    historico.save(request)
                    log(u'No considerar creditos: %s - %s - %s' % (record.inscripcion, record, motivo), request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'validar':
            try:
                record = RecordAcademico.objects.get(pk=request.POST['id'])
                f = ConsiderarForm(request.POST)
                if f.is_valid():
                    motivo = f.cleaned_data['motivo']
                    record.valida = True
                    record.save(request)
                    historico = record.mi_historico()
                    historico.valida = record.valida
                    historico.save(request)
                    log(u'Considerar creditos: %s - %s - %s' % (record.inscripcion, record, motivo), request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'novalidarpromedio':
            try:
                record = RecordAcademico.objects.get(pk=request.POST['id'])
                f = ConsiderarForm(request.POST)
                if f.is_valid():
                    motivo = f.cleaned_data['motivo']
                    record.validapromedio = False
                    record.save(request)
                    historico = record.mi_historico()
                    historico.validapromedio = record.validapromedio
                    historico.save(request)
                    log(u'No considerar promedio: %s - %s - %s' % (record.inscripcion, record, motivo), request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'validarpromedio':
            try:
                record = RecordAcademico.objects.get(pk=request.POST['id'])
                f = ConsiderarForm(request.POST)
                if f.is_valid():
                    motivo = f.cleaned_data['motivo']
                    record.validapromedio = True
                    record.save(request)
                    historico = record.mi_historico()
                    historico.validapromedio = record.validapromedio
                    historico.save(request)
                    log(u'Considerar promedio: %s - %s - %s' % (record.inscripcion, record, motivo), request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'adddocumento':
            try:
                form = DocumentoInscripcionForm(request.POST, request.FILES)
                inscripcion = Inscripcion.objects.get(pk=request.POST['inscripcion'])
                if form.is_valid():
                    nfile = request.FILES['archivo']
                    nfile._name = generar_nombre("documentos_", nfile._name)
                    archivo = Archivo(nombre=form.cleaned_data['nombre'],
                                      fecha=datetime.now().date(),
                                      archivo=nfile,
                                      tipo_id=ARCHIVO_TIPO_GENERAL,
                                      inscripcion=inscripcion)
                    archivo.save(request)
                    log(u'Adiciono documento de inscripcion: %s - %s' % (inscripcion, archivo), request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'importar':
            try:
                form = ImportarArchivoXLSForm(request.POST, request.FILES)
                if form.is_valid():
                    hoy = datetime.now().date()
                    nfile = request.FILES['archivo']
                    nfile._name = generar_nombre("importacion_", nfile._name)
                    archivo = Archivo(nombre='IMPORTACION INSCRIPCIONES',
                                      fecha=datetime.now().date(),
                                      archivo=nfile,
                                      tipo_id=ARCHIVO_TIPO_GENERAL)
                    archivo.save(request)
                    workbook = xlrd.open_workbook(archivo.archivo.file.name)
                    sheet = workbook.sheet_by_index(0)
                    linea = 1
                    for rowx in range(sheet.nrows):
                        puntosalva = transaction.savepoint()
                        try:
                            cols = sheet.row_values(rowx)
                            cedula = cols[0].strip().upper()
                            bandera = True
                            if Persona.objects.filter(cedula=cols[0]).exists():
                                bandera = False
                            if bandera:
                                if cedula and Persona.objects.filter(cedula=cedula).exists():
                                    return JsonResponse({"result": "bad", "mensaje": u"Existe una persona con esta identificación: %s." % cedula})
                                persona = Persona(cedula=cedula,
                                                  apellido1=cols[1],
                                                  apellido2=cols[2],
                                                  nombres=cols[3],
                                                  sexo_id=int(cols[4]),
                                                  nacimiento=xlrd.xldate.xldate_as_datetime(cols[5], workbook.datemode).date(),
                                                  provincianac=cols[6],
                                                  cantonnac=cols[7],
                                                  ciudad=cols[8],
                                                  email=cols[9],
                                                  telefono_conv=cols[10],
                                                  telefono=cols[11])
                                persona.save(request)
                                username = calculate_username(persona)
                                usuario = generar_usuario(persona, username, ALUMNOS_GROUP_ID)
                                if EMAIL_INSTITUCIONAL_AUTOMATICO:
                                    persona.emailinst = username + '@' + EMAIL_DOMAIN
                                    persona.save(request)
                                grupo = None
                                if UTILIZA_GRUPOS_ALUMNOS:
                                    grupo = Grupo.objects.get(pk=int(cols[15]))
                                    carrera = grupo.carrera
                                    sesion = grupo.sesion
                                    modalidad = grupo.modalidad
                                    sede = grupo.sede
                                else:
                                    sesion = Sesion.objects.get(pk=int(cols[12]))
                                    carrera = Carrera.objects.get(pk=int(cols[13]))
                                    modalidad = Modalidad.objects.get(pk=int(cols[14]))
                                    sede = Sede.objects.get(pk=int(cols[15]))
                            else:
                                persona = Persona.objects.filter(cedula=cols[0])[0]
                                sesion = Sesion.objects.get(pk=int(cols[12]))
                                carrera = Carrera.objects.get(pk=int(cols[13]))
                                modalidad = Modalidad.objects.get(pk=int(cols[14]))
                                sede = Sede.objects.get(pk=int(cols[15]))
                            inscripcion = Inscripcion(persona=persona,
                                                      fecha=xlrd.xldate.xldate_as_datetime(cols[18], workbook.datemode).date(),
                                                      carrera=carrera,
                                                      modalidad=modalidad,
                                                      sesion=sesion,
                                                      sede=sede,
                                                      colegio=cols[16])
                            inscripcion.save(request)
                            persona.crear_perfil(inscripcion=inscripcion)
                            documentos = DocumentosDeInscripcion(inscripcion=inscripcion,
                                                                 titulo=False,
                                                                 acta=False,
                                                                 cedula=False,
                                                                 votacion=False,
                                                                 actaconv=False,
                                                                 partida_nac=False,
                                                                 pre=False,
                                                                 observaciones_pre='',
                                                                 fotos=False)
                            documentos.save(request)
                            preguntasinscripcion = inscripcion.preguntas_inscripcion()
                            inscripcion.persona.mi_perfil()
                            inscripciontesdrive = InscripcionTesDrive(inscripcion=inscripcion,
                                                                      licencia=False,
                                                                      record=False,
                                                                      certificado_tipo_sangre=False,
                                                                      prueba_psicosensometrica=False,
                                                                      certificado_estudios=False)
                            inscripciontesdrive.save(request)
                            # inscripcion.mi_malla()
                            inscripcion.malla_inscripcion()
                            inscripcion.actualizar_nivel()
                            if USA_TIPOS_INSCRIPCIONES:
                                inscripciontipoinscripcion = InscripcionTipoInscripcion(inscripcion=inscripcion,
                                                                                        tipoinscripcion_id=TIPO_INSCRIPCION_INICIAL)
                                inscripciontipoinscripcion.save(request)
                            log(u'Adiciono inscripcion: %s' % inscripcion, request, "add")
                            log(u'Importo inscripcion: %s' % inscripcion, request, "add")
                            # persona.creacion_persona(request.session['nombresistema'])
                            linea += 1
                            print(linea)
                            transaction.savepoint_commit(puntosalva)
                        except Exception as ex:
                            transaction.savepoint_rollback(puntosalva)
                            return JsonResponse({"result": "bad", "mensaje": u"Error al ingresar la linea: %s" % linea})
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == "generarenvio":
            try:
                listapersonainscripcion = Inscripcion.objects.filter(modalidad_id=3, confimacion_online=False,envioemail=False)[:40]
                for personainscripcion in listapersonainscripcion:
                    personainscripcion.envio_correo_admision(20, request)
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                pass

        elif action == 'fechainicioconvalidacion':
            try:
                form = FechaInicioConvalidacionInscripcionForm(request.POST)
                inscripcion = Inscripcion.objects.get(pk=request.POST['id'])
                if form.is_valid():
                    inscripcion.fechainicioconvalidacion = form.cleaned_data['fecha']
                    inscripcion.save(request)
                    log(u'Adiciono fecha inicio convalidacion: %s' % inscripcion, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'fechainicioprimernivel':
            try:
                form = FechaInicioPrimerNivelInscripcionForm(request.POST)
                inscripcion = Inscripcion.objects.get(pk=int(request.POST['id']))
                if form.is_valid():
                    inscripcion.fechainicioprimernivel = form.cleaned_data['fecha']
                    inscripcion.save(request)
                    log(u'Adiciono fecha inicio primer nivel: %s' % inscripcion, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'deldocumento':
            try:
                archivo = Archivo.objects.get(pk=request.POST['id'])
                inscripcion = archivo.inscripcion
                archivo.delete()
                log(u'Elimino documento de inscripcion: %s - %s' % (inscripcion, archivo), request, "del")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'asignaturas':
            try:
                inscripcion = Inscripcion.objects.get(pk=request.POST['id'])
                if int(request.POST['tipo']) == 1:
                    asignaturas = Asignatura.objects.filter(id__in=[x.asignatura.id for x in inscripcion.malla_inscripcion().malla.asignaturamalla_set.all()])
                else:
                    asignaturas = Asignatura.objects.all()
                return JsonResponse({"result": "ok", 'listado': [(x.id, x.nombre) for x in asignaturas]})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'activar':
            try:
                inscripcion = Inscripcion.objects.get(pk=request.POST['id'])
                ui = inscripcion.persona.usuario
                ui.is_active = True
                ui.save()
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'desactivar':
            try:
                inscripcion = Inscripcion.objects.get(pk=request.POST['id'])
                ui = inscripcion.persona.usuario
                ui.is_active = False
                ui.save()
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'desactivarperfil':
            try:
                inscripcion = Inscripcion.objects.get(pk=request.POST['id'])
                inscripcion.activo = False
                inscripcion.save(request)
                log(u'Desactivo perfil de usuario: %s' % inscripcion, request, "edit")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'activarperfil':
            try:
                inscripcion = Inscripcion.objects.get(pk=request.POST['id'])
                inscripcion.activo = True
                inscripcion.save(request)
                log(u'Activo perfil de usuario: %s' % inscripcion, request, "edit")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'buscarpreinscripcion':
            try:
                if PreInscrito.objects.filter(cedula=request.POST['ced']).exists():
                    preinscrito = PreInscrito.objects.filter(cedula=request.POST['ced'])[0]
                    return JsonResponse({'result': 'ok', 'preinscrito': unicode(preinscrito), 'preinscrito_id': preinscrito.id})
                else:
                    return JsonResponse({'result': 'bad'})
            except Exception as ex:
                return JsonResponse({'result': 'bad', "mensaje": u'Error al obtener los datos'})

        elif action == 'buscarpersona':
            try:
                if Persona.objects.filter(cedula=request.POST['ced']).exists():
                    persona = Persona.objects.get(cedula=request.POST['ced'])
                    return JsonResponse({'result': 'ok', 'apellido1': persona.apellido1, 'apellido2': persona.apellido2, 'nombres': persona.nombres, 'sexo': persona.sexo_id})
                else:
                    return JsonResponse({'result': 'bad'})
            except Exception as ex:
                return JsonResponse({'result': 'bad', "mensaje": u'Error al obtener los datos'})

        elif action == 'resetear':
            try:
                inscripcion = Inscripcion.objects.get(pk=request.POST['id'])
                resetear_clave(inscripcion.persona)
                log(u'Reseteo clave de inscripcion: %s' % inscripcion, request, "add")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'resetear_clave_admision_virtual':
            try:
                inscripcion = Inscripcion.objects.get(pk=request.POST['id'])
                if inscripcion.carrera.mi_coordinacion2() == 9:
                    resetear_clave_admision_ldap(inscripcion.persona)
                    # from moodle import moodle
                    # bestudiante = moodle.BuscarUsuario(periodo, 2, 'idnumber', inscripcion.persona.identificacion())
                    # if not bestudiante:
                    #     bestudiante = moodle.BuscarUsuario(periodo, 2, 'idnumber', inscripcion.persona.identificacion())
                    # if bestudiante['users']:
                    #     if 'auth' in bestudiante['users'][0]:
                    #         auth = bestudiante['users'][0]['auth']
                    #         if auth == 'ldap':
                    #             resetear_clave_admision_ldap(inscripcion.persona)
                    #         else:
                    #             resetear_clave_admision_manual(inscripcion.persona)
                else:
                    resetear_clave(inscripcion.persona)
                log(u'Reseteo clave de inscripcion : %s' % inscripcion, request, "add")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'retirocarrera':
            try:
                inscripcion = Inscripcion.objects.get(pk=request.POST['id'])
                form = RetiradoMateriaForm(request.POST)
                if form.is_valid():
                    if not inscripcion.retirocarrera_set.exists():
                        retiro = RetiroCarrera(inscripcion=inscripcion,
                                               fecha=datetime.now().date(),
                                               motivo=form.cleaned_data['motivo'])
                        retiro.save(request)
                    log(u'Retiro de carrera: %s' % inscripcion, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'addadministrativo':
            try:
                inscripcion = Inscripcion.objects.get(pk=request.POST['id'])
                if inscripcion.persona.es_administrador():
                    return JsonResponse({"result": "bad", "mensaje": u"Ya existe un perfil administrativo para este usuario."})
                administrativo = Administrativo(persona=inscripcion.persona,
                                                contrato='',
                                                fechaingreso=datetime.now().date())
                administrativo.save(request)
                g = Group.objects.get(pk=variable_valor('ADMINISTRATIVOS_GROUP_ID'))
                g.user_set.add(inscripcion.persona.usuario)
                g.save(request)
                inscripcion.persona.crear_perfil(administrativo=administrativo)
                log(u'Adiciono personal administrativo: %s' % administrativo, request, "add")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'adddocente':
            try:
                inscripcion = Inscripcion.objects.get(pk=request.POST['id'])
                if inscripcion.persona.es_profesor():
                    return JsonResponse({"result": "bad", "mensaje": u"Ya existe un perfil de docente para este usuario."})
                profesor = Profesor(persona=inscripcion.persona,
                                    activo=True,
                                    fechaingreso=datetime.now().date(),
                                    coordinacion=Coordinacion.objects.all()[0],
                                    dedicacion=TiempoDedicacionDocente.objects.all()[0])
                profesor.save(request)
                grupo = Group.objects.get(pk=PROFESORES_GROUP_ID)
                grupo.user_set.add(inscripcion.persona.usuario)
                grupo.save(request)
                inscripcion.persona.crear_perfil(profesor=profesor)
                log(u'Adiciono profesor: %s' % profesor, request, "add")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'infoasignatura':
            try:
                inscripcion = Inscripcion.objects.get(pk=request.POST['iid'])
                malla = inscripcion.malla_inscripcion().malla
                if malla.asignaturamalla_set.filter(asignatura__id=request.POST['id']).exists():
                    asignaturamalla = malla.asignaturamalla_set.filter(asignatura__id=request.POST['id'])[0]
                    return JsonResponse({'result': 'ok', 'creditos': asignaturamalla.creditos, 'horas': asignaturamalla.horas})
                else:
                    asignatura = Asignatura.objects.get(pk=request.POST['id'])
                    return JsonResponse({'result': 'ok', 'creditos': asignatura.creditos, 'horas': 0})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'imprimirporcentajeporcarrera':
            try:
                data['title'] = u'Porcentaje de Actividades Generales'
                listapersonainscripcion = Inscripcion.objects.filter(modalidad_id=3)
                if int(request.POST['carrera']) > 0:
                    listapersonainscripcion = listapersonainscripcion.filter(carrera_id=int(request.POST['carrera']))
                    data['carrera'] = Carrera.objects.get(pk=int(request.POST['carrera'])).nombre_completo()
                if int(request.POST['pais']) > 0:
                    listapersonainscripcion = listapersonainscripcion.filter(persona__pais_id=int(request.POST['pais']))
                    data['pais']=  Pais.objects.get(pk=int(request.POST['pais'])).nombre
                if request.POST['tipo'] and request.POST['tipo'] == '1':
                    listapersonainscripcion = listapersonainscripcion.filter(persona__ppl=True)
                    data['ppl'] = 'PPL'
                data['listapersonainscripcion'] = listapersonainscripcion

                return conviert_html_to_pdf('inscripciones_admision/imprimirporcentajeporcarrera.html',
                                            {'pagesize': 'A4',
                                             'data': data,
                                             'listapersonainscripcion': listapersonainscripcion,
                                             })
            except Exception as ex:
                pass

        elif action == 'verificarestadomatricula':
            try:
                form = ImportarArchivoXLSForm(request.POST, request.FILES)
                if form.is_valid():
                    nfile = request.FILES['archivo']
                    nfile._name = generar_nombre("importacion_", nfile._name)
                    archivo = Archivo(nombre='IMPORTACION ALUMNOS ONLINE',
                                      fecha=datetime.now().date(),
                                      archivo=nfile,
                                      tipo_id=ARCHIVO_TIPO_GENERAL)
                    archivo.save()
                    a = 0
                    miarchivo = openpyxl.load_workbook(archivo.archivo.file.name)
                    lista = miarchivo.get_sheet_by_name('Hoja1')
                    totallista = lista.rows
                    for filas in totallista:
                        a += 1
                        if a > 1:
                            cedula = filas[0].value.strip().upper()
                            inscripcion = None
                            inscrito = 'NO MATRICULADO'
                            estadoactual = 0
                            if filas[4].value == None:
                                if Persona.objects.filter(Q(cedula=cedula) | Q(pasaporte=cedula), status=True).exists():
                                    perso = Persona.objects.filter(Q(cedula=cedula) | Q(pasaporte=cedula), status=True)[0]
                                    carrera = filas[3].value
                                    estadoactual = 0
                                    inscrito = 'NO MATRICULADO'
                                    if Inscripcion.objects.filter(persona=perso, carrera_id=carrera):
                                        inscrito = 'MATRICULADO'
                                        inscripcion = Inscripcion.objects.filter(persona=perso, carrera_id=carrera)[0]
                                        sqluser = "select distinct a.USER_ID as idusuario, a.EID as cedula, b.LAST_NAME as apellido, b.FIRST_NAME, b.EMAIL as email from SAKAI_USER_ID_MAP as a inner join SAKAI_USER as b on b.USER_ID=a.USER_ID inner join SAKAI_SITE_USER as c on c.USER_ID=a.USER_ID where a.EID='" + str(inscripcion.persona.cedula) + "'"
                                        resultados_usuario = querymysqlsakai(sqluser, True)
                                        id_usuario = resultados_usuario[0][0]
                                        sql = "select * from(select site.SITE_ID as sitio, site.TITLE as nombre_curso, site.type as tipo, " \
                                              "(select SRR.ROLE_NAME from SAKAI_REALM_RL_GR SRRG " \
                                              "inner join SAKAI_REALM SR on SRRG.REALM_KEY = SR.REALM_KEY " \
                                              "inner join SAKAI_REALM_ROLE SRR on SRRG.ROLE_KEY = SRR.ROLE_KEY  " \
                                              "where SR.REALM_ID = CONCAT ( '/site/', site.SITE_ID) and SRRG.USER_ID = c.USER_ID and SRRG.ACTIVE = '1') as rol " \
                                              "from SAKAI_USER_ID_MAP as a " \
                                              "inner join SAKAI_USER as b on b.USER_ID=a.USER_ID " \
                                              "inner join SAKAI_SITE_USER as c on c.USER_ID=a.USER_ID " \
                                              "inner join SAKAI_SITE as site on site.SITE_ID = c.SITE_ID " \
                                              "where a.EID= '" + str(inscripcion.persona.cedula) + "' and site.IS_SOFTLY_DELETED = 0 " \
                                                                                                   "and site.PUBVIEW = '1') as roles " \
                                                                                                   "where rol = 'Student' order by nombre_curso"
                                        resultados_cursos = querymysqlsakai(sql, True)
                                        for x in resultados_cursos:
                                            if estadoactual == 1:
                                                break
                                            sql = "select tareas.ASSIGNMENT_ID from ASSIGNMENT_ASSIGNMENT tareas where tareas.CONTEXT = '" + str(x[0]) + "'"
                                            resultados_tareas = querymysqlsakai(sql, True)
                                            for tarea in resultados_tareas:
                                                sqlenviotarea = "select submission.XML " \
                                                                "from ASSIGNMENT_SUBMISSION submission " \
                                                                "inner join SAKAI_REALM_RL_GR SRRG on SRRG.USER_ID = submission.SUBMITTER_ID " \
                                                                "INNER JOIN SAKAI_REALM SR ON SRRG.REALM_KEY = SR.REALM_KEY " \
                                                                "INNER JOIN SAKAI_REALM_ROLE SRR ON SRRG.ROLE_KEY = SRR.ROLE_KEY " \
                                                                "where submission.CONTEXT ='" + tarea[0] + "' and SR.REALM_ID = CONCAT ('/site/', '" + x[0] + "') " \
                                                                                                                                                              "AND SRRG.ACTIVE = '1'and  SRR.ROLE_NAME = 'Student' and SRRG.USER_ID='" + id_usuario + "'"
                                                resultados_enviotarea = querymysqlsakai(sqlenviotarea, True)
                                                if resultados_enviotarea:
                                                    id_xml_tarea = resultados_enviotarea[0][0]
                                                    xmltext = minidom.parseString(id_xml_tarea)
                                                    itemlist = xmltext.getElementsByTagName("submission")
                                                    datesubmitted = itemlist[0].getAttribute('datesubmitted')
                                                    if datesubmitted:
                                                        estadoactual = 1
                                                        break
                                filas[4].value = inscrito
                                if estadoactual == 1:
                                    filas[5].value = 'EN CURSO'
                                else:
                                    filas[5].value = 'RETIRADO'
                                miarchivo.save("lista_inscritos.xlsx")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'imprimirtabla2pdf':
            try:
                return conviert_html_to_pdf('inscripciones_admision/imprimirtabla2pdf.html',
                                            {'pagesize': 'A4',
                                             'data': data,
                                             'lista_carreras':lista_carreras,
                                             'periodo':periodo,'cantidad_graduados_online':cantidad_graduados_online
                                             })
            except Exception as ex:
                pass

        elif action == 'imprimirtabla3pdf':
            try:
                return conviert_html_to_pdf('inscripciones_admision/imprimirtabla3pdf.html',
                                            {'pagesize': 'A4',
                                             'data': data,
                                             'lista_carreras':lista_carreras,
                                             'periodo':periodo,'cantidad_graduados_online':cantidad_graduados_online
                                             })
            except Exception as ex:
                pass

        elif action == 'imprimirtabla4pdf':
            try:
                cantidad_aprobados=Matricula.objects.filter(status=True, nivel__periodo=periodo, aprobado=True).count()
                cantidad_reprobados=Matricula.objects.filter(status=True, nivel__periodo=periodo, aprobado=False,inscripcion__modalidad=3).exclude(inscripcion__desertaonline=True).count()
                cantidad_destaron=Matricula.objects.filter(status=True, nivel__periodo=periodo,inscripcion__desertaonline=True).count()
                return conviert_html_to_pdf('inscripciones_admision/imprimirtabla4pdf.html',
                                            {'pagesize': 'A4',
                                             'data': data,
                                             'lista_carreras':lista_carreras,
                                             'cantidad_aprobados':cantidad_aprobados,
                                             'cantidad_reprobados':cantidad_reprobados,
                                             'cantidad_destaron':cantidad_destaron,
                                             'periodo':periodo, 'cantidad_graduados_online':cantidad_graduados_online
                                             })
            except Exception as ex:
                pass

        elif action == 'imprimirtabla2excel':
            try:
                __author__ = 'Unemi'
                style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
                style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                style1 = easyxf(num_format_str='D-MMM-YY')
                font_style = XFStyle()
                font_style.font.bold = True
                font_style2 = XFStyle()
                font_style2.font.bold = False
                wb = Workbook(encoding='utf-8')
                ws = wb.add_sheet('tabla2')
                ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                ws.write(1, 0, "SECCIÓN DE ADMISIÓN Y NIVELACIÓN", font_style2)
                ws.write(2, 0, "TABLA 2. NÚMERO ESTUDIANTES MATRICULADOS.", font_style2)
                ws.write(3, 0,"NIVELACIÓN DE CARRERA MODALIDA EN LÍNEA SEGUNDO SEMESTRE 2018", font_style2)
                response = HttpResponse(content_type="application/ms-excel")
                response['Content-Disposition'] = 'attachment; filename=tabla2' + random.randint(1, 10000).__str__() + '.xls'
                columns = [
                    (u"No.", 1000),
                    (u"CARRERA", 9000),
                    (u"MATRIZ DE TERCER NIVEL", 6000),
                    (u"MATRICULADOS", 6000),
                ]
                row_num = 4
                for col_num in range(len(columns)):
                    ws.write(row_num, col_num, columns[col_num][0], font_style)
                    ws.col(col_num).width = columns[col_num][1]
                date_format = xlwt.XFStyle()
                date_format.num_format_str = 'yyyy/mm/dd'
                row_num = 5
                i = 0
                for r in lista_carreras:
                    i += 1
                    campo1 = i
                    campo2 = r.nombre_completo()
                    campo3 = r.cantidad_graduados()
                    campo4 = r.matriculados(periodo)
                    ws.write(row_num, 0, campo1, font_style2)
                    ws.write(row_num, 1, campo2, font_style2)
                    ws.write(row_num, 2, campo3, font_style2)
                    ws.write(row_num, 3, campo4, font_style2)
                    row_num += 1
                ws.write(row_num, 1, "Total", font_style2)
                ws.write(row_num, 2, "---", font_style2)
                ws.write(row_num, 3, cantidad_graduados_online, font_style2)
                wb.save(response)
                return response
            except Exception as ex:
                pass

        elif action == 'imprimirtabla3excel':
            try:
                __author__ = 'Unemi'
                style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
                style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                style1 = easyxf(num_format_str='D-MMM-YY')
                font_style = XFStyle()
                font_style.font.bold = True
                font_style2 = XFStyle()
                font_style2.font.bold = False
                wb = Workbook(encoding='utf-8')
                ws = wb.add_sheet('tabla2')
                ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                ws.write(1, 0, "SECCIÓN DE ADMISIÓN Y NIVELACIÓN", font_style2)
                ws.write(2, 0, "TABLA 3. RESUMEN DE ESTUDIANTES EFECTIVAMENTE MATRICULADOS.", font_style2)
                ws.write(3, 0, "NIVELACIÓN DE CARRERA MODALIDA EN LÍNEA SEGUNDO SEMESTRE 2018", font_style2)
                response = HttpResponse(content_type="application/ms-excel")
                response['Content-Disposition'] = 'attachment; filename=tabla2' + random.randint(1, 10000).__str__() + '.xls'
                columns = [
                    (u"No.", 1000),
                    (u"CARRERA", 9000),
                    (u"MATRIZ DE TERCER NIVEL", 6000),
                    (u"MATRICULADOS", 6000),
                    (u"ASISTEN", 6000),
                    (u"RETIRAN", 6000),
                ]
                row_num = 4
                for col_num in range(len(columns)):
                    ws.write(row_num, col_num, columns[col_num][0], font_style)
                    ws.col(col_num).width = columns[col_num][1]
                date_format = xlwt.XFStyle()
                date_format.num_format_str = 'yyyy/mm/dd'
                row_num = 5
                i = 0
                for r in lista_carreras:
                    i += 1
                    campo1 = i
                    campo2 = r.nombre_completo()
                    campo3 = '24'
                    campo4 = r.matriculados(periodo)
                    campo5 = '123'
                    campo6 = '123'
                    ws.write(row_num, 0, campo1, font_style2)
                    ws.write(row_num, 1, campo2, font_style2)
                    ws.write(row_num, 2, campo3, font_style2)
                    ws.write(row_num, 3, campo4, font_style2)
                    ws.write(row_num, 4, campo5, font_style2)
                    ws.write(row_num, 5, campo6, font_style2)
                    row_num += 1
                ws.write(row_num, 1, "Total", font_style2)
                ws.write(row_num, 2, "---", font_style2)
                ws.write(row_num, 3, cantidad_graduados_online, font_style2)
                ws.write(row_num, 4, "Total", font_style2)
                ws.write(row_num, 5, "Total", font_style2)
                wb.save(response)
                return response
            except Exception as ex:
                pass

        elif action == 'imprimirtabla4excel':
            try:
                __author__ = 'Unemi'
                style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
                style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                style1 = easyxf(num_format_str='D-MMM-YY')
                font_style = XFStyle()
                font_style.font.bold = True
                font_style2 = XFStyle()
                font_style2.font.bold = False
                wb = Workbook(encoding='utf-8')
                ws = wb.add_sheet('tabla2')
                ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                ws.write(1, 0, "SECCIÓN DE ADMISIÓN Y NIVELACIÓN", font_style2)
                ws.write(2, 0, "TABLA 4. RESUMEN DE RESULTADOS DE ESTUDIANTES MATRICULADOS.", font_style2)
                ws.write(3, 0,"NIVELACIÓN DE CARRERA MODALIDA EN LÍNEA SEGUNDO SEMESTRE 2018", font_style2)
                response = HttpResponse(content_type="application/ms-excel")
                response['Content-Disposition'] = 'attachment; filename=tabla2' + random.randint(1, 10000).__str__() + '.xls'
                columns = [
                    (u"No.", 1000),
                    (u"CARRERA", 9000),
                    (u"MATRIZ DE TERCER NIVEL", 6000),
                    (u"MATRICULADOS", 6000),
                    (u"APROBADOS", 6000),
                    (u"REPROBADOS", 6000),
                    (u"RETIRADOS", 6000),
                ]
                row_num = 4
                for col_num in range(len(columns)):
                    ws.write(row_num, col_num, columns[col_num][0], font_style)
                    ws.col(col_num).width = columns[col_num][1]
                date_format = xlwt.XFStyle()
                date_format.num_format_str = 'yyyy/mm/dd'
                row_num = 5
                i = 0
                for r in lista_carreras:
                    i += 1
                    campo1 = i
                    campo2 = r.nombre_completo()
                    campo3 = '24'
                    campo4 = r.matriculados(periodo)
                    campo5 = '123'
                    campo6 = '123'
                    campo7 = '5346'
                    ws.write(row_num, 0, campo1, font_style2)
                    ws.write(row_num, 1, campo2, font_style2)
                    ws.write(row_num, 2, campo3, font_style2)
                    ws.write(row_num, 3, campo4, font_style2)
                    ws.write(row_num, 4, campo5, font_style2)
                    ws.write(row_num, 5, campo6, font_style2)
                    ws.write(row_num, 6, campo7, font_style2)
                    row_num += 1
                ws.write(row_num, 1, "Total", font_style2)
                ws.write(row_num, 2, "---", font_style2)
                ws.write(row_num, 3, cantidad_graduados_online, font_style2)
                ws.write(row_num, 4, "Total", font_style2)
                ws.write(row_num, 5, "Total", font_style2)
                ws.write(row_num, 6, "Total", font_style2)
                wb.save(response)
                return response
            except Exception as ex:
                pass

        elif action == 'excel_resumen_sobre70':
            try:
                __author__ = 'Unemi'
                idcarrera = 0
                idmateria = 0
                if 'idcarrera' in request.POST:
                    idcarrera = int(request.POST['idcarrera'])
                if 'idmateria' in request.POST:
                    idmateria = int(request.POST['idmateria'])
                carrera = None
                materia = None
                matriculas = None
                nombre = ''
                if idcarrera > 0:
                    carrera = Carrera.objects.get(id=idcarrera)
                    matriculas = carrera.matriculados_efectivo(periodo)
                    nombre = str(carrera.nombre_completo())
                if idmateria > 0:
                    materia = Materia.objects.get(id=idmateria)
                    lista = MateriaAsignada.objects.values_list('matricula__id', flat=True).filter(status=True, materia=materia)
                    matriculas = Matricula.objects.filter(status=True, inscripcion__carrera=carrera,nivel__periodo=periodo, id__in=lista).distinct().order_by('inscripcion__persona__apellido1','inscripcion__persona__apellido2','inscripcion__persona__nombres')
                    nombre = nombre + " / " + str(materia.nombre_completo())
                style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
                style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                style1 = easyxf(num_format_str='D-MMM-YY')
                font_style = XFStyle()
                font_style.font.bold = True
                font_style2 = XFStyle()
                font_style2.font.bold = False
                wb = Workbook(encoding='utf-8')
                ws = wb.add_sheet('resumen')
                ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                ws.write(1, 0, "SECCIÓN DE ADMISIÓN Y NIVELACIÓN", font_style2)
                ws.write(2, 0, "RESUMEN CALIFICACIONES PORCENTAJE - " + nombre, font_style2)
                response = HttpResponse(content_type="application/ms-excel")
                response['Content-Disposition'] = 'attachment; filename=resumen' + random.randint(1, 10000).__str__() + '.xls'
                columns = [
                    (u"Rango.", 2000),
                    (u"Cantidad", 5000),
                ]
                row_num = 4
                for col_num in range(len(columns)):
                    ws.write(row_num, col_num, columns[col_num][0], font_style)
                    ws.col(col_num).width = columns[col_num][1]
                date_format = xlwt.XFStyle()
                date_format.num_format_str = 'yyyy/mm/dd'
                row_num = 5
                campo2=0
                campo4=0
                campo6=0
                campo8=0
                campo10=0
                campo12=0
                campo14=0
                for matricula in matriculas:
                    porcentaje = 0
                    try:
                        if idmateria > 0:
                            porcentaje = matricula.inscripcion.porcentaje_actividades_por_asignatura(materia.codigosakai)
                        elif idcarrera > 0:
                            porcentaje = matricula.inscripcion.porcentaje_total_estudiante()
                    except Exception as ex:
                        porcentaje = 0

                    if porcentaje <= 10:
                        campo2 += 1
                    elif porcentaje >= 10.01 and porcentaje <= 20:
                        campo4 += 1
                    elif porcentaje >= 20.01 and porcentaje <= 30:
                        campo6 += 1
                    elif porcentaje >= 30.01 and porcentaje <= 40:
                        campo8 += 1
                    elif porcentaje >= 40.01 and porcentaje <= 50:
                        campo10 += 1
                    elif porcentaje >= 50.01 and porcentaje <= 60:
                        campo12 += 1
                    elif porcentaje >= 60.01 and porcentaje <= 70:
                        campo14 += 1

                campo1 = '0 - 10%'
                campo3 = '10.01 - 20%'
                campo5 = '20.01 - 30%'
                campo7 = '30.01 - 40%'
                campo9 = '40.01 - 50%'
                campo11 = '50.01 - 60%'
                campo13 = '60.01 - 70%'
                ws.write(row_num, 0, campo1, font_style2)
                ws.write(row_num, 1, campo2, font_style2)
                ws.write(row_num+1, 0, campo3, font_style2)
                ws.write(row_num+1, 1, campo4, font_style2)
                ws.write(row_num+2, 0, campo5, font_style2)
                ws.write(row_num+2, 1, campo6, font_style2)
                ws.write(row_num+3, 0, campo7, font_style2)
                ws.write(row_num+3, 1, campo8, font_style2)
                ws.write(row_num+4, 0, campo9, font_style2)
                ws.write(row_num+4, 1, campo10, font_style2)
                ws.write(row_num+5, 0, campo11, font_style2)
                ws.write(row_num+5, 1, campo12, font_style2)
                ws.write(row_num+6, 0, campo13, font_style2)
                ws.write(row_num+6, 1, campo14, font_style2)
                wb.save(response)
                return response
            except Exception as ex:
                pass

        elif action == 'excel_resumen_alumnos':
            try:
                __author__ = 'Unemi'
                idcarrera = 0
                idmateria = 0
                if 'idcarrera' in request.POST:
                    idcarrera = int(request.POST['idcarrera'])
                if 'idmateria' in request.POST:
                    idmateria = int(request.POST['idmateria'])
                carrera = None
                materia = None
                matriculas = None
                nombre=''
                if idcarrera > 0:
                    carrera = Carrera.objects.get(id=idcarrera)
                    matriculas = carrera.matriculados_efectivo(periodo)
                    nombre = str(carrera.nombre_completo())
                if idmateria > 0:
                    materia = Materia.objects.get(id=idmateria)
                    lista = MateriaAsignada.objects.values_list('matricula__id', flat=True).filter(status=True, materia=materia)
                    matriculas = Matricula.objects.filter(status=True, inscripcion__carrera=carrera, nivel__periodo=periodo, id__in=lista).distinct().order_by('inscripcion__persona__apellido1','inscripcion__persona__apellido2','inscripcion__persona__nombres')
                    nombre = nombre + " / "+str(materia.nombre_completo())
                style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
                style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                style1 = easyxf(num_format_str='D-MMM-YY')
                font_style = XFStyle()
                font_style.font.bold = True
                font_style2 = XFStyle()
                font_style2.font.bold = False
                wb = Workbook(encoding='utf-8')
                ws = wb.add_sheet('resumen')
                ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                ws.write(1, 0, "SECCIÓN DE ADMISIÓN Y NIVELACIÓN", font_style2)
                ws.write(2, 0, "RESUMEN CALIFICACIONES PORCENTAJES - " + nombre, font_style2)
                response = HttpResponse(content_type="application/ms-excel")
                response['Content-Disposition'] = 'attachment; filename=resumen' + random.randint(1, 10000).__str__() + '.xls'
                columns = [
                    (u"Cedula .", 2000),
                    (u"Alumno .", 2000),
                    (u"Porcentaje", 5000),
                ]
                row_num = 4
                for col_num in range(len(columns)):
                    ws.write(row_num, col_num, columns[col_num][0], font_style)
                    ws.col(col_num).width = columns[col_num][1]
                date_format = xlwt.XFStyle()
                date_format.num_format_str = 'yyyy/mm/dd'
                row_num = 6
                campo1 = ''
                campo2=0

                for matricula in matriculas:
                    campo3 = str(matricula.inscripcion.persona.identificacion())
                    campo1 = str(matricula.inscripcion.persona.nombre_completo_inverso())
                    try:
                        if idmateria > 0:
                            campo2 = matricula.inscripcion.porcentaje_actividades_por_asignatura(materia.codigosakai)
                        elif idcarrera > 0:
                            campo2 = matricula.inscripcion.porcentaje_total_estudiante()
                    except Exception as ex:
                        campo2 = 0

                    # try:
                    #     campo2 =  matricula.inscripcion.porcentaje_total_estudiante()
                    # except Exception as ex:
                    #     campo2 = 0
                    ws.write(row_num, 0, campo3, font_style2)
                    ws.write(row_num, 1, campo1, font_style2)
                    ws.write(row_num, 2, campo2, font_style2)
                    row_num += 1
                wb.save(response)
                return response
            except Exception as ex:
                pass

        elif action == 'excel_resumen_asignatura':
            try:
                __author__ = 'Unemi'
                idcarrera = request.POST['idcarrera']
                idmateria = request.POST['idmateria']
                carrera = Carrera.objects.get(id=idcarrera)
                lista = []
                materia = Materia.objects.get(id=idmateria)
                style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
                style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                title = easyxf(
                    'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                style1 = easyxf(num_format_str='D-MMM-YY')
                font_style = XFStyle()
                font_style.font.bold = True
                font_style2 = XFStyle()
                font_style2.font.bold = False
                wb = Workbook(encoding='utf-8')
                ws = wb.add_sheet('resumen')
                ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                ws.write(1, 0, "SECCIÓN DE ADMISIÓN Y NIVELACIÓN", font_style2)
                ws.write(2, 0, "Resumen calificaciones porcentajes", font_style2)
                response = HttpResponse(content_type="application/ms-excel")
                response['Content-Disposition'] = 'attachment; filename=resumen' + random.randint(1,10000).__str__() + '.xls'
                columns = [
                    (u"Rango.", 2000),
                    (u"Cantidad", 5000),
                ]
                row_num = 4
                for col_num in range(len(columns)):
                    ws.write(row_num, col_num, columns[col_num][0], font_style)
                    ws.col(col_num).width = columns[col_num][1]
                date_format = xlwt.XFStyle()
                date_format.num_format_str = 'yyyy/mm/dd'
                row_num = 5
                campo1 = '0 - 10%'
                campo2 = materia.calcular_cantidad_alumnos(0, 10)
                campo3 = '10.01 - 20%'
                campo4 = materia.calcular_cantidad_alumnos(10.01, 20)
                campo5 = '20.01 - 30%'
                campo6 = materia.calcular_cantidad_alumnos(20.01, 30)
                campo7 = '30.01 - 40%'
                campo8 = materia.calcular_cantidad_alumnos(30.01, 40)
                campo9 = '40.01 - 50%'
                campo10 = materia.calcular_cantidad_alumnos(40.01, 50)
                campo11 = '50.01 - 60%'
                campo12 = materia.calcular_cantidad_alumnos(50.01, 60)
                campo13 = '60.01 - 70%'
                campo14 = materia.calcular_cantidad_alumnos(60.01, 70)
                ws.write(row_num, 0, campo1, font_style2)
                ws.write(row_num, 1, campo2, font_style2)
                ws.write(row_num + 1, 0, campo3, font_style2)
                ws.write(row_num + 1, 1, campo4, font_style2)
                ws.write(row_num + 2, 0, campo5, font_style2)
                ws.write(row_num + 2, 1, campo6, font_style2)
                ws.write(row_num + 3, 0, campo7, font_style2)
                ws.write(row_num + 3, 1, campo8, font_style2)
                ws.write(row_num + 4, 0, campo9, font_style2)
                ws.write(row_num + 4, 1, campo10, font_style2)
                ws.write(row_num + 5, 0, campo11, font_style2)
                ws.write(row_num + 5, 1, campo12, font_style2)
                ws.write(row_num + 6, 0, campo13, font_style2)
                ws.write(row_num + 6, 1, campo14, font_style2)
                wb.save(response)
                return response
            except Exception as ex:
                pass

        elif action == 'pdf_resumen_sobre70':
            try:
                idcarrera = 0
                idmateria = 0
                uno = 0
                dos = 0
                tres = 0
                cuatro = 0
                cinco = 0
                seis = 0
                siete = 0
                if 'idcarrera' in request.POST:
                    idcarrera = int(request.POST['idcarrera'])
                if 'idmateria' in request.POST:
                    idmateria = int(request.POST['idmateria'])
                carrera = None
                materia = None
                matriculas = None
                nombre = ''
                if idcarrera > 0:
                    carrera = Carrera.objects.get(id=idcarrera)
                    matriculas = carrera.matriculados_efectivo(periodo)
                    nombre = str(carrera.nombre_completo())
                if idmateria > 0:
                    materia = Materia.objects.get(id=idmateria)
                    lista = MateriaAsignada.objects.values_list('matricula__id', flat=True).filter(status=True, materia=materia)
                    matriculas = Matricula.objects.filter(status=True, inscripcion__carrera=carrera,nivel__periodo=periodo, id__in=lista).distinct().order_by('inscripcion__persona__apellido1','inscripcion__persona__apellido2','inscripcion__persona__nombres')
                    nombre = nombre + " / " + str(materia.nombre_completo())
                for matricula in matriculas:
                    porcentaje = 0
                    try:
                        if idmateria > 0:
                            porcentaje = matricula.inscripcion.porcentaje_actividades_por_asignatura(materia.codigosakai)
                        elif idcarrera > 0:
                            porcentaje = matricula.inscripcion.porcentaje_total_estudiante
                    except Exception as ex:
                        porcentaje = 0

                    if porcentaje <= 0 and 0 == 0:
                        uno += 1
                    elif porcentaje >= 0 and porcentaje <= 10:
                        uno += 1
                    elif porcentaje >= 10.01 and porcentaje <= 20:
                        dos += 1
                    elif porcentaje >= 20.01 and porcentaje <= 30:
                        tres += 1
                    elif porcentaje >= 30.01 and porcentaje <= 40:
                        cuatro += 1
                    elif porcentaje >= 40.01 and porcentaje <= 50:
                        cinco += 1
                    elif porcentaje >= 50.01 and porcentaje <= 60:
                        seis += 1
                    elif porcentaje >= 60.01 and porcentaje <= 70:
                        siete += 1
                return conviert_html_to_pdf('inscripciones_admision/pdf_resumen_sobre70.html',
                                            {'pagesize': 'A4',
                                             'data': data,
                                             'nombre': nombre,
                                             'uno': uno,
                                             'dos': dos,
                                             'tres': tres,
                                             'cuatro': cuatro,
                                             'cinco': cinco,
                                             'seis': seis,
                                             'siete': siete
                                             })
            except Exception as ex:
                pass

        elif action == 'pdf_resumen_alumno':
            try:
                idcarrera = 0
                idmateria = 0
                if 'idcarrera' in request.POST:
                    idcarrera = int(request.POST['idcarrera'])
                if 'idmateria' in request.POST:
                    idmateria = int(request.POST['idmateria'])
                carrera = None
                materia = None
                matriculas = None
                nombre = ''
                if idcarrera > 0:
                    carrera = Carrera.objects.get(id=idcarrera)
                    matriculas = carrera.matriculados_efectivo(periodo)
                    nombre = str(carrera.nombre_completo())
                if idmateria > 0:
                    materia = Materia.objects.get(id=idmateria)
                    lista = MateriaAsignada.objects.values_list('matricula__id', flat=True).filter(status=True,materia=materia)
                    matriculas = Matricula.objects.filter(status=True, inscripcion__carrera=carrera,nivel__periodo=periodo, id__in=lista).distinct().order_by('inscripcion__persona__apellido1','inscripcion__persona__apellido2','inscripcion__persona__nombres')
                    nombre = nombre + " / " + str(materia.nombre_completo())
                return conviert_html_to_pdf('inscripciones_admision/pdf_resumen_alumno.html',
                                            {'pagesize': 'A4',
                                             'data': data,
                                             'nombre': nombre,
                                             'matriculas': matriculas,
                                             'idmateria':idmateria,
                                             'idcarrera':idcarrera,'materia':materia
                                             })
            except Exception as ex:
                pass

        elif action == 'seguimiento_estudiantes':
            try:
                data['inscripcion']= inscripcion = Inscripcion.objects.get(pk=request.POST['idi'])
                if 'materia' in request.POST:
                    materias = Materia.objects.get(pk=int(request.POST['materia']))
                if 'fini' in request.POST:
                    data['fini'] = fini = request.POST['fini'] if request.POST['fini'] else None
                if 'ffin' in request.POST:
                    data['ffin'] = ffin = request.POST['ffin'] if request.POST['ffin'] else None
                if 'tipo' in request.POST:
                    data['tipo']= tipo = request.POST['tipo'] if request.POST['tipo'] else None

                data['listaactividades'] = inscripcion.seguimiento_actividades(materias.codigosakai,fini,ffin,tipo)
                template = get_template("inscripciones_admision/seguimiento_actividades.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'html': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'seguimiento_actividades':
            try:
                if 'idreferencia' in request.POST:
                    referencia = request.POST['idreferencia']
                    id_actividad = referencia.split("/")[5]
                inscripcion = Inscripcion.objects.get(pk=request.POST['idi'])
                data['listaactividades'] = inscripcion.seguimiento_actividad(id_actividad)
                template = get_template("inscripciones_admision/seguimiento_actividad.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'data': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'excel_resumen_ponderacion':
            try:
                __author__ = 'Unemi'
                idcarrera = 0
                idmateria = 0
                carrera = None
                matriculas = None
                porcentaje_calificacion = 0
                nombre = ''
                materia = None
                if 'idcarrera' in request.POST:
                    idcarrera = int(request.POST['idcarrera'])
                if 'idmateria' in request.POST:
                    idmateria = int(request.POST['idmateria'])

                if idcarrera > 0:
                    carrera = Carrera.objects.get(id=idcarrera)
                    nombre = str(carrera.nombre_completo())
                if idmateria > 0:
                    materia = Materia.objects.get(id=idmateria)
                    lista = MateriaAsignada.objects.values_list('matricula__id', flat=True).filter(status=True,
                                                                                                   materia=materia)
                    matriculas = Matricula.objects.filter(status=True, inscripcion__carrera=carrera,
                                                          nivel__periodo=periodo, id__in=lista).distinct()
                    porcentaje_calificacion = matriculas[0].inscripcion.porcentaje_equivalente_asignatura(materia.id)
                    nombre = nombre + " " + str(materia.nombre_completo())

                style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
                style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                title = easyxf(
                    'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                style1 = easyxf(num_format_str='D-MMM-YY')
                font_style = XFStyle()
                font_style.font.bold = True
                font_style2 = XFStyle()
                font_style2.font.bold = False
                wb = Workbook(encoding='utf-8')
                ws = wb.add_sheet('resumen')
                ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                ws.write(1, 0, "SECCIÓN DE ADMISIÓN Y NIVELACIÓN", font_style2)
                ws.write(2, 0, "PORCENTAJE DE CALIFICACIÓN POR ASGINATURA SEGÚN SU PONDERACIÓN - ", font_style2)
                ws.write(3, 0, nombre + ' Equivalente al ' + str(porcentaje_calificacion) + '%', font_style2)
                response = HttpResponse(content_type="application/ms-excel")
                response['Content-Disposition'] = 'attachment; filename=resumen' + random.randint(1,10000).__str__() + '.xls'
                columns = [
                    (u"Rango.", 2000),
                    (u"Cantidad.", 5000),
                ]
                row_num = 5
                for col_num in range(len(columns)):
                    ws.write(row_num, col_num, columns[col_num][0], font_style)
                    ws.col(col_num).width = columns[col_num][1]
                date_format = xlwt.XFStyle()
                date_format.num_format_str = 'yyyy/mm/dd'
                row_num = 6

                rango1 = 0
                rango2 = 0
                rango3 = 0
                campo2 = 0
                campo4 = 0
                campo6 = 0
                campo1 = ''
                campo3 = ''
                campo5 = ''

                rango1 = materia.rango_porcentaje()[0]
                rango2 = materia.rango_porcentaje()[1]
                if materia.rango_porcentaje().__len__() == 3:
                    rango3 = materia.rango_porcentaje()[2]

                campo1 = '0% - ' + str(rango1) + '%'
                campo3 = str(rango1 + 0.01) + '% - ' + str(rango2) + '%'
                campo5 = str(rango2 + 0.01) + '% - ' + str(rango3) + '%' if rango3 > 0 else ''

                for matricula in matriculas:
                    porcentaje = matricula.inscripcion.porcentaje_por_asignatura(materia.codigosakai)
                    if porcentaje <= 0:
                        campo2 += 1
                    elif porcentaje >= 0 and porcentaje <= rango1:
                        campo2 += 1
                    elif porcentaje >= (rango1 + 0.01) and porcentaje <= rango2:
                        campo4 += 1
                    elif rango3 > 0:
                        if porcentaje >= (rango2 + 0.01) and porcentaje <= rango3:
                            campo6 += 1

                ws.write(row_num, 0, campo1, font_style2)
                ws.write(row_num, 1, campo2, font_style2)
                ws.write(row_num + 1, 0, campo3, font_style2)
                ws.write(row_num + 1, 1, campo4, font_style2)
                if not campo5 == '':
                    ws.write(row_num + 2, 0, campo5, font_style2)
                    ws.write(row_num + 2, 1, campo6, font_style2)
                ws.write(row_num + 3, 0, "Total Estudiantes", font_style2)
                ws.write(row_num + 3, 1, matriculas.count(), font_style2)
                wb.save(response)
                return response
            except Exception as ex:
                pass

        elif action == 'imprimir_pdf_resumen_ponderacion':
            try:
                data['title'] = u'Porcentaje Ponderacion Asignatura'
                idcarrera = 0
                idmateria = 0
                carrera = None
                matriculas = None
                nombre=''
                materia=None
                rango1 = 0
                rango2 = 0
                rango3 = 0
                con1 =0
                con2 =0
                con3 =0
                lista_porcentaje=[]
                if 'idcarrera' in request.POST:
                    idcarrera = int(request.POST['idcarrera'])
                if 'idmateria' in request.POST:
                    idmateria = int(request.POST['idmateria'])

                if idcarrera > 0:
                    carrera = Carrera.objects.get(id=idcarrera)
                    nombre = str(carrera.nombre_completo())
                if idmateria > 0:
                    materia = Materia.objects.get(id=idmateria)
                    lista = MateriaAsignada.objects.values_list('matricula__id', flat=True).filter(status=True, materia=materia)
                    matriculas = Matricula.objects.filter(status=True, inscripcion__carrera=carrera,nivel__periodo=periodo, id__in=lista).distinct()
                    nombre = nombre + " " + str(materia.nombre_completo())
                    porcentaje_calificacion = matriculas[0].inscripcion.porcentaje_equivalente_asignatura(materia.id)

                rango1 = materia.rango_porcentaje()[0]
                rango2 = materia.rango_porcentaje()[1]
                if materia.rango_porcentaje().__len__() == 3:
                    rango3 = materia.rango_porcentaje()[2]

                for matricula in matriculas:
                    porcentaje = matricula.inscripcion.porcentaje_por_asignatura(materia.codigosakai)
                    if porcentaje <=0:
                        con1+=1
                    elif porcentaje >= 0 and porcentaje <=rango1:
                        con1+=1
                    elif porcentaje >= (rango1+0.01) and porcentaje<= rango2:
                        con2+=1
                    elif rango3 > 0:
                        if porcentaje >= (rango2+0.01) and porcentaje <= rango3:
                            con3+=1

                lista_porcentaje.append((str(0) + '% - ' + str(rango1) + '%', con1))
                lista_porcentaje.append((str(rango1+0.01)+'% - ' + str(rango2)+'%',con2))
                if rango3 > 0:
                    lista_porcentaje.append((str(rango2+0.01)+'% - ' + str(rango3)+'%',con3))

                return conviert_html_to_pdf('inscripciones_admision/imprimir_ponderacion_asignatura.html',
                                            {'pagesize': 'A4',
                                             'data': data,
                                             'materia': materia,
                                             'nombre': nombre,
                                             'porcentaje_calificacion': porcentaje_calificacion,
                                             'lista_porcentaje': lista_porcentaje
                                             })
            except Exception as ex:
                pass

        elif action == 'excel_resumen_completo':
            try:
                __author__ = 'Unemi'
                if 'idcarrera' in request.POST:
                    idcarrera = int(request.POST['idcarrera'])
                    carrera = Carrera.objects.get(id=idcarrera)
                matriculas = Matricula.objects.filter(status=True, nivel__periodo=periodo, inscripcion__modalidad_id=3, inscripcion__carrera=carrera).distinct().order_by('inscripcion__persona__apellido1','inscripcion__persona__apellido2','inscripcion__persona__nombres')
                style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
                style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                style1 = easyxf(num_format_str='D-MMM-YY')
                font_style = XFStyle()
                font_style.font.bold = True
                font_style2 = XFStyle()
                font_style2.font.bold = False
                wb = Workbook(encoding='utf-8')
                ws = wb.add_sheet('resumen70general')
                ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                ws.write(1, 0, "SECCIÓN DE ADMISIÓN Y NIVELACIÓN", font_style2)
                ws.write(2, 0, "RESUMEN CALIFICACIONES PORCENTAJE 70 GENERAL" , font_style2)
                response = HttpResponse(content_type="application/ms-excel")
                response['Content-Disposition'] = 'attachment; filename=resumen70general' + random.randint(1, 10000).__str__() + '.xls'
                columns = [
                    (u"Carrera", 2000),
                    (u"Asignatura", 5000),
                    (u"Paralelo", 5000),
                    (u"Rango", 5000),
                    (u"Cantidad", 5000),
                ]
                row_num = 4
                for col_num in range(len(columns)):
                    ws.write(row_num, col_num, columns[col_num][0], font_style)
                    ws.col(col_num).width = columns[col_num][1]
                date_format = xlwt.XFStyle()
                date_format.num_format_str = 'yyyy/mm/dd'
                row_num = 5
                campo2 = 0
                campo4 = 0
                campo6 = 0
                campo8 = 0
                campo10 = 0
                campo12 = 0
                campo14 = 0
                campo1 = '0 - 10%'
                campo3 = '10.01 - 20%'
                campo5 = '20.01 - 30%'
                campo7 = '30.01 - 40%'
                campo9 = '40.01 - 50%'
                campo11 = '50.01 - 60%'
                campo13 = '60.01 - 70%'
                # for carrera in lista_carreras:
                materiaasignada = MateriaAsignada.objects.values_list('materia__id',flat=True).filter(status=True, matricula__nivel__periodo=periodo,
                                                                                                      matricula__inscripcion__modalidad_id=3,
                                                                                                      matricula__inscripcion__carrera=carrera).distinct()
                materias = Materia.objects.filter(status=True, id__in=materiaasignada)
                for materia in materias:
                    lista = MateriaAsignada.objects.values_list('matricula__id',flat=True).filter(status=True, matricula__nivel__periodo=periodo,
                                                                                                  matricula__inscripcion__modalidad_id=3,
                                                                                                  matricula__inscripcion__carrera=carrera, materia=materia).distinct()
                    matriculas = matriculas.filter(id__in=lista)
                    ws.write(row_num, 0, carrera.nombre_completo(), font_style2)
                    ws.write(row_num, 1, materia.nombre_completo(), font_style2)
                    ws.write(row_num, 2, materia.paralelo, font_style2)
                    campo2 = 0
                    campo4 = 0
                    campo6 = 0
                    campo8 = 0
                    campo10 = 0
                    campo12 = 0
                    campo14 = 0
                    for matricula in matriculas:
                        porcentaje = 0
                        try:
                            porcentaje = matricula.inscripcion.porcentaje_actividades_por_asignatura(materia.codigosakai)
                            # porcentaje = 0
                        except Exception as ex:
                            porcentaje = 0

                        if porcentaje <= 10:
                            campo2 += 1
                        elif porcentaje >= 10.01 and porcentaje <= 20:
                            campo4 += 1
                        elif porcentaje >= 20.01 and porcentaje <= 30:
                            campo6 += 1
                        elif porcentaje >= 30.01 and porcentaje <= 40:
                            campo8 += 1
                        elif porcentaje >= 40.01 and porcentaje <= 50:
                            campo10 += 1
                        elif porcentaje >= 50.01 and porcentaje <= 60:
                            campo12 += 1
                        elif porcentaje >= 60.01 and porcentaje <= 70:
                            campo14 += 1
                    ws.write(row_num, 3, campo1, font_style2)
                    ws.write(row_num, 4, campo2, font_style2)
                    ws.write(row_num+1, 3, campo3, font_style2)
                    ws.write(row_num+1, 4, campo4, font_style2)
                    ws.write(row_num+2, 3, campo5, font_style2)
                    ws.write(row_num+2, 4, campo6, font_style2)
                    ws.write(row_num+3, 3, campo7, font_style2)
                    ws.write(row_num+3, 4, campo8, font_style2)
                    ws.write(row_num+4, 3, campo9, font_style2)
                    ws.write(row_num+4, 4, campo10, font_style2)
                    ws.write(row_num+5, 3, campo11, font_style2)
                    ws.write(row_num+5, 4, campo12, font_style2)
                    ws.write(row_num+6, 3, campo13, font_style2)
                    ws.write(row_num+6, 4, campo14, font_style2)
                    row_num+=7
                wb.save(response)
                return response
            except Exception as ex:
                pass

        elif action== 'migrar_examen_moodle':
            try:
                form = ImportarArchivoXLSForm(request.POST, request.FILES)
                if form.is_valid():
                    nfile = request.FILES['archivo']
                    nfile._name = generar_nombre("importacion_", nfile._name)
                    archivo = Archivo(nombre='IMPORTACION EXAMEN',
                                      fecha=datetime.now().date(),
                                      archivo=nfile,
                                      tipo_id=ARCHIVO_TIPO_GENERAL)
                    archivo.save()
                    workbook = xlrd.open_workbook(archivo.archivo.file.name)
                    sheet = workbook.sheet_by_index(0)
                    linea = 1
                    # periodo = Periodo.objects.get(pk=80)
                    modalidad = Modalidad.objects.get(pk=3)
                    for rowx in range(sheet.nrows):
                        if linea > 1:
                            cols = sheet.row_values(rowx)
                            identificacion = str(cols[0]).strip().upper() if cols[0] else None
                            asignatura = int(cols[1]) if cols[1] else None
                            calificacion = null_to_decimal(cols[8], 2) if cols[8] else None
                            if calificacion and asignatura and identificacion:
                                if Persona.objects.filter(Q(cedula=identificacion) | Q(pasaporte=identificacion)).exists():
                                    persona = \
                                    Persona.objects.filter(Q(cedula=identificacion) | Q(pasaporte=identificacion))[0]
                                    if Inscripcion.objects.filter(status=True, persona=persona,
                                                                  modalidad=modalidad).exists():
                                        inscripcion = \
                                        Inscripcion.objects.filter(status=True, persona=persona, modalidad=modalidad)[0]
                                        if inscripcion.matricula_set.filter(nivel__periodo=periodo).exists():
                                            matricula = inscripcion.matricula_set.filter(nivel__periodo=periodo)

                                            materiaasignada = \
                                            MateriaAsignada.objects.filter(status=True, materiaasignadaretiro__isnull=True,
                                                                           materia__asignatura_id=asignatura,
                                                                           matricula=matricula)[0]

                                            if not ActividadesSakaiAlumno.objects.filter(tipo=4, status=True,
                                                                                         inscripcion=inscripcion,
                                                                                         materia=materiaasignada.materia).exists():
                                                actividadguardada = ActividadesSakaiAlumno(inscripcion=inscripcion,
                                                                                           materia=materiaasignada.materia,
                                                                                           nombreactividadsakai="Examen",
                                                                                           tipo=4,
                                                                                           nota=calificacion if calificacion else 0,
                                                                                           notaposible=30
                                                                                           )
                                                actividadguardada.save()
                                            else:
                                                actividadguardada = \
                                                ActividadesSakaiAlumno.objects.filter(tipo=4, status=True,
                                                                                      inscripcion=inscripcion,
                                                                                      materia=materiaasignada.materia)[0]
                                                if calificacion > actividadguardada.nota:
                                                    actividadguardada.nota = calificacion
                                                    actividadguardada.notaposible = 30
                                                    actividadguardada.save()
                        linea += 1


                # if 'carr' in request.POST:
                #     idcarrera = int(request.POST['carr'])
                #     carrera = Carrera.objects.get(id=idcarrera)
                # matriculas = Matricula.objects.filter(status=True, nivel__periodo=periodo, inscripcion__modalidad__id=3, inscripcion__carrera=carrera).distinct().order_by('inscripcion__persona__apellido1', 'inscripcion__persona__apellido2', 'inscripcion__persona__nombres')
                # for matricula in matriculas:
                #     for materiaasignada in matricula.materiaasignada_set.filter(status=True):
                #         try:
                #             calificacion = materiaasignada.materia.examen_moodle(matricula.inscripcion.persona)
                #         except Exception as ex:
                #             calificacion = 0
                #         if calificacion.__len__() > 0:
                #             if not ActividadesSakaiAlumno.objects.filter(tipo=4,status=True, idactividadsakai=calificacion[0][0], inscripcion=matricula.inscripcion, materia=materiaasignada.materia).exists():
                #                 actividadguardada = ActividadesSakaiAlumno(inscripcion=matricula.inscripcion,
                #                                                            materia=materiaasignada.materia,
                #                                                            idactividadsakai=calificacion[0][0],
                #                                                            nombreactividadsakai=calificacion[0][2],
                #                                                            tipo=4, nota=calificacion[0][1]if calificacion.__len__() > 0 else 0,
                #                                                            notaposible=30
                #                                                            )
                #                 actividadguardada.save(request)
                #             else:
                #                 if calificacion.__len__()>0:
                #                     actividadguardada = ActividadesSakaiAlumno.objects.filter(tipo=4,status=True, idactividadsakai=calificacion[0][0], inscripcion=matricula.inscripcion, materia=materiaasignada.materia)[0]
                #                     actividadguardada.nota = calificacion[0][1] if calificacion.__len__() > 0 else 0
                #                     actividadguardada.notaposible = 30
                #                     actividadguardada.save(request)

                log(u'Migración de datos de examenes : %s' % personasesion, request, "add")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'migrar_actividades':
            materror = None
            try:
                if 'carr' in request.POST:
                    idcarrera = int(request.POST['carr'])
                    carrera = Carrera.objects.get(id=idcarrera)
                matriculas = Matricula.objects.filter(status=True, nivel__periodo=periodo, inscripcion__carrera=carrera).distinct().order_by('inscripcion__persona__apellido1', 'inscripcion__persona__apellido2', 'inscripcion__persona__nombres')
                for matricula in matriculas:
                    materror = matricula
                    if periodo.usa_sakai:
                        matricula.migrar_gestion(periodo)
                    elif periodo.usa_moodle:
                        matricula.migrar_gestion_moodle(periodo)
                    log(u'Migración de datos de actividades: %s' % personasesion, request, "add")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.seguardart_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos. %s %s" % (ex, materror)})

        elif action == 'editmallahistorica':
            try:
                recordacademico = RecordAcademico.objects.get(pk=request.POST['id'])
                f = MallaHistoricaForm(request.POST)
                if f.is_valid():
                    recordacademico.asignaturamallahistorico = f.cleaned_data['asignaturamallahistorico']
                    recordacademico.asignaturamalla = recordacademico.inscripcion.asignatura_en_asignaturamalla(f.cleaned_data['asignaturamallahistorico'].asignatura )
                    recordacademico.save(request)
                    return JsonResponse({"result": "ok"})
                else:
                     raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'asignacion_cupos':
            try:
                if 'carr' in request.POST:
                    idcarrera = int(request.POST['carr'])
                    carrera = Carrera.objects.get(id=idcarrera)
                Matricula.objects.filter(inscripcion__carrera=carrera,nivel__periodo=periodo,materiaasignada__materia__asignaturamalla__malla__carrera__coordinacion__id=9,inscripcion__modalidad_id=3).update(aprobado=False)
                # carreras = Carrera.objects.filter(malla__asignaturamalla__materia__nivel__periodo=periodo,coordinacion__id=9, modalidad=3).distinct()
                # for carrera in carreras:
                matriculas = Matricula.objects.filter(status=True, nivel__periodo=periodo,materiaasignada__materia__asignaturamalla__malla__carrera__coordinacion__id=9, inscripcion__modalidad__id=3, inscripcion__carrera=carrera).distinct().order_by('promedionotasvirtual')
                #     matriculas.update(aprobado=False)
                listaconcuposlimpios = []
                for matricula in matriculas:
                    if matricula.promedionotasvirtual >= 70 and matricula.verificar_todas_materias_aprobadas(periodo):
                        listaconcuposlimpios.append(matricula.id)
                Matricula.objects.filter(id__in=listaconcuposlimpios).update(aprobado=True)
                log(u'Asignacion de Cupos: %s' % personasesion, request, "add")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos. %s %s"})

        elif action == 'actualizar_datos_reporte':
            try:
                inscripcion=Inscripcion.objects.get(pk=request.POST['id'])
                campo1 = str(inscripcion.persona.identificacion())
                campo2 = str(inscripcion.carrera.nombre_completo())
                campo3 = str(inscripcion.persona.apellido1 + " " + inscripcion.persona.apellido2)
                campo4 = str(inscripcion.persona.nombres)
                campo5 = str(inscripcion.persona.email)
                campo6 = str(inscripcion.persona.emailinst)
                campo7 = str(inscripcion.persona.telefono)
                campo8 = str(inscripcion.persona.pais) if inscripcion.persona.pais else ""
                campo9 = str(inscripcion.persona.provincia) if inscripcion.persona.provincia else ""
                campo10 = str(inscripcion.persona.canton) if inscripcion.persona.canton else ""
                campo11 = str(inscripcion.persona.parroquia) if inscripcion.persona.parroquia else ""
                campo12 = str("SI") if inscripcion.persona.tiene_ficha_confirmada() else "NO"
                try:
                    campo13 = inscripcion.porcentaje_total_estudiante()
                except Exception as ex:
                    campo13 = 0
                campo14 = str(inscripcion.persona.mi_perfil().tipodiscapacidad) if inscripcion.persona.mi_perfil().tipodiscapacidad else "S/N"
                campo15 = str(inscripcion.persona.mi_perfil().porcientodiscapacidad) if inscripcion.persona.mi_perfil().porcientodiscapacidad else "0"
                campo16 = str(inscripcion.persona.mi_perfil().carnetdiscapacidad) if inscripcion.persona.mi_perfil().carnetdiscapacidad else "S/N"
                campo19 = "SI" if inscripcion.desertaonline else "NO"
                campo20 = "SI" if inscripcion.persona.ppl else "NO"
                campo21 = str(inscripcion.persona.pais.id) if inscripcion.persona.pais else "S/N"
                campo22 = str(inscripcion.persona.provincia.id) if inscripcion.persona.provincia else "S/N"
                campo23 = str(inscripcion.persona.canton.id) if inscripcion.persona.canton else "S/N"

                mismaterias = inscripcion.asignaturas_sakai()
                for materia in mismaterias:
                    campo17 = materia[1]
                    try:
                        campo18 = inscripcion.porcentaje_actividades_por_asignatura(materia[0])
                    except Exception as ex:
                        campo18 = 0

                    try:
                        campo25 = inscripcion.porcentaje_por_asignatura(materia[0])
                    except Exception as ex:
                        campo25 = 0

                    try:
                        campo26 = inscripcion.porcentaje_equivalente_asignatura(materia[0])
                    except Exception as ex:
                        campo26 = 0
                    try:
                        sql = "select id_ficha_est from ficha_estudiante where cedula='" + str(campo1) + "' and asignatura_paralelo='" + str(materia[1]) + "'"
                        resultados_enviotarea = querymysqlconsulta(sql, True)
                        if resultados_enviotarea:
                            sql = "update ficha_estudiante set cedula='" + str(campo1) + "', apel='" + str(campo3) + "', " \
                                                                                                                     "nom='" + str(campo4) + "',ppl='" + str(campo20) + "',email_per='" + str(campo5) + "', email_inst='" + str(campo6) + "'," \
                                                                                                                                                                                                                                          " celular='" + str(campo7) + "', parroquia='" + str(campo11) + "', canton='" + str(campo10) + "', provincia='" + str(campo9) + "', pais='" + str(campo8) + "',confirmo_ficha='" + str(campo12) + "', discapacidad='" + str(campo14) + "',porcent_disc='" + str(campo15) + "', deserto='" + str(campo19) + "',porcentaje_global= " + str(campo13) + ", pocentaje_ponderacion=" + str(campo25) + ", ponderacion=" + str(campo26) + ", porcentaje_asignatura=" + str(campo18) + " where cedula='" + str(campo1) + "' and asignatura_paralelo='" + str(materia[1]) + "'"
                        else:
                            sql = " INSERT INTO `ficha_estudiante` ( `cedula`, `apel`, `nom`, " \
                                  "`ppl`, `email_per`, `email_inst`, " \
                                  "`celular`, `pais`, `provincia`, `canton`, `parroquia`,  `confirmo_ficha`, " \
                                  "`discapacidad`, `porcent_disc`, `deserto`, " \
                                  "`porcentaje_global`, `porcentaje_asignatura`, `pocentaje_ponderacion`, `ponderacion`, " \
                                  "`asignatura_paralelo`) " \
                                  "VALUES ( '" + str(campo1) + "', '" + str(campo3) + "', '" + str(campo2) + "', '" \
                                  + str(campo20) + "', '" + str(campo5) + "', '" + str(campo6) + "', '" \
                                  + str(campo7) + "', '" + str(campo8) + "', '" + str(campo9) + "', '" + str(campo10) + "', '"+ str(campo11)+ "', '"+ str(campo12) + "', '" \
                                  + str(campo14) + "', '" + str(campo15) + "', '" + str(campo19) + "', " \
                                  + str(campo13) + ", " + str(campo18) + ", " + str(campo25) + ", "+ str(campo26) + ", '" + str(materia[1]) + "');"
                        querymysqlconsulta(sql)
                    except Exception as ex:
                        pass
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        # elif action == 'modificarnota':
        #     try:
        #         actividad = ActividadesSakaiAlumno.objects.get(pk=request.POST['mid'])
        #         valor = request.POST['vc']
        #         actividad.nota = valor
        #         actividad.save(request)
        #         return JsonResponse({'result': 'ok', 'valor': actividad.nota})
        #     except Exception as ex:
        #         transaction.set_rollback(True)
        #         return JsonResponse({'result': 'bad',"mensaje": u"Error al actualizar nota."})
        #
        # elif action == 'modificarnotaposible':
        #     try:
        #         actividad = ActividadesSakaiAlumno.objects.get(pk=request.POST['mid'])
        #         valor = request.POST['vc']
        #         actividad.notaposible = valor
        #         actividad.save(request)
        #         return JsonResponse({'result': 'ok', 'valor': actividad.notaposible})
        #     except Exception as ex:
        #         transaction.set_rollback(True)
        #         return JsonResponse({'result': 'bad',"mensaje": u"Error al actualizar codigo sakai."})
        #
        # elif action == 'delactividadsakai':
        #     try:
        #         actividad = ActividadesSakaiAlumno.objects.get(pk=request.POST['id'])
        #         # log(u'Elimino actividad misgrada del sakai: %s' % actividad, request, "del")
        #         actividad.delete()
        #         return JsonResponse({"result": "ok"})
        #     except Exception as ex:
        #         transaction.set_rollback(True)
        #         return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'cupos_carreras':
            try:
                __author__ = 'Unemi'

                style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
                style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                title = easyxf(
                    'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                style1 = easyxf(num_format_str='D-MMM-YY')
                font_style = XFStyle()
                font_style.font.bold = True
                font_style2 = XFStyle()
                font_style2.font.bold = False
                wb = Workbook(encoding='utf-8')
                ws = wb.add_sheet('listado_docente_online')
                ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                response = HttpResponse(content_type="application/ms-excel")
                response['Content-Disposition'] = 'attachment; filename=listado_docente_online' + random.randint(1,
                                                                                                                 10000).__str__() + '.xls'
                columns = [
                    (u"Carrera .", 5000),
                    (u"Cantidad Aprobados", 5000),
                ]
                row_num = 4
                for col_num in range(len(columns)):
                    ws.write(row_num, col_num, columns[col_num][0], font_style)
                    ws.col(col_num).width = columns[col_num][1]
                date_format = xlwt.XFStyle()
                date_format.num_format_str = 'yyyy/mm/dd'
                row_num = 5

                aprobados = Matricula.objects.filter(status=True, nivel__periodo=periodo, aprobado=True).count()
                for carrera in Carrera.objects.filter(modalidad=3):
                    matriculas = Matricula.objects.filter(status=True, nivel__periodo=periodo, aprobado=True,inscripcion__carrera=carrera).count()
                    ws.write(row_num, 0, carrera.nombre_completo(), font_style2)
                    ws.write(row_num, 1, matriculas, font_style2)
                    row_num += 1

                ws.write(row_num, 0, "Total Aprobado", font_style2)
                ws.write(row_num, 1, aprobados, font_style2)
                wb.save(response)
                return response
            except Exception as ex:
                pass

        elif action == 'alumnos_carreras':
            try:
                __author__ = 'Unemi'

                style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
                style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                title = easyxf(
                    'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                style1 = easyxf(num_format_str='D-MMM-YY')
                font_style = XFStyle()
                font_style.font.bold = True
                font_style2 = XFStyle()
                font_style2.font.bold = False
                wb = Workbook(encoding='utf-8')
                ws = wb.add_sheet('listado_docente_online')
                ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                response = HttpResponse(content_type="application/ms-excel")
                response['Content-Disposition'] = 'attachment; filename=listado_alumnos_carreras' + random.randint(1,10000).__str__() + '.xls'
                columns = [
                    (u"Carrera .", 5000),
                    (u"Cantidad Alumnos", 5000),
                ]
                row_num = 4
                for col_num in range(len(columns)):
                    ws.write(row_num, col_num, columns[col_num][0], font_style)
                    ws.col(col_num).width = columns[col_num][1]
                date_format = xlwt.XFStyle()
                date_format.num_format_str = 'yyyy/mm/dd'
                row_num = 5

                for carrera in Carrera.objects.filter(modalidad=3):
                    matriculas = Matricula.objects.filter(status=True, nivel__periodo=periodo, estado_matricula__in=[2,3],inscripcion__carrera=carrera).count()
                    ws.write(row_num, 0, carrera.nombre_completo(), font_style2)
                    ws.write(row_num, 1, matriculas, font_style2)
                    row_num += 1
                wb.save(response)
                return response
            except Exception as ex:
                pass

        elif action == 'migrar_actividades_alumno':
            try:
                if 'inscripcion' in request.POST:
                    inscripcion = int(request.POST['inscripcion'])
                    inscripcion = Inscripcion.objects.get(id=inscripcion)
                    matriculas = Matricula.objects.filter(status=True, nivel__periodo=periodo,inscripcion=inscripcion)
                    actividadguardada = None
                    for matricula in matriculas:
                        if periodo.usa_sakai:
                            matricula.migrar_gestion(periodo)
                        else:
                            matricula.migrar_gestion_moodle(periodo)
                    log(u'Migración de datos de actividades sakai de un alumno: %s %s' % (personasesion, inscripcion), request, "add")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos. %s" % (ex) })

        elif action == 'migrar_examen_moodle_carrera':
            try:
                if 'carr' in request.POST:
                    idcarrera = int(request.POST['carr'])
                    carrera = Carrera.objects.get(id=idcarrera)
                    matriculas = Matricula.objects.filter(status=True, nivel__periodo=periodo, inscripcion__carrera=carrera).distinct().order_by('inscripcion__persona__apellido1', 'inscripcion__persona__apellido2', 'inscripcion__persona__nombres')
                    for matricula in matriculas:
                        lista = ""
                        sql_examenes = ""
                        nota = 0
                        fullname = ""
                        shortname = ""
                        materiaasignada = matricula.materiaasignada_set.filter(status=True, materiaasignadaretiro__isnull=True).order_by('materia__id').distinct()
                        for x in materiaasignada:
                            if x.materia.idcursomoodle != materiaasignada.order_by('-materia__id')[0].materia.idcursomoodle:
                                if x.materia.idcursomoodle:
                                    lista += str(x.materia.idcursomoodle) + ","
                            else:
                                lista += str(x.materia.idcursomoodle)
                        cursor = connections['db_moodle_virtual'].cursor()
                        sql_examenes = """
                                                SELECT DISTINCT co.fullname, co.shortname, notatest.grade
                                                FROM mooc_quiz_grades notatest
                                                INNER JOIN mooc_quiz test ON notatest.quiz=test.id
                                                INNER JOIN mooc_user u ON u.id=notatest.userid
                                                INNER JOIN mooc_course co ON  co.id=test.course
                                                WHERE notatest.grade >= 0 AND u.idnumber='%s' AND not co.id IN (%s)
                                            """ % (str(matricula.inscripcion.persona.identificacion()), lista)
                        cursor.execute(sql_examenes)
                        datosexamenes = cursor.fetchall()
                        for dato in datosexamenes:
                            nota = dato[2] if dato else 0
                            fullname = dato[0] if dato else ""
                            shortname = dato[1] if dato else ""
                            if nota:
                                if MateriaAsignada.objects.filter(Q(materia__asignatura__nombre__icontains=fullname) | Q(
                                        materia__asignatura__nombre__icontains=shortname),
                                                                  status=True, materiaasignadaretiro__isnull=True,
                                                                  matricula=matricula).exists():
                                    materiaasignada = MateriaAsignada.objects.filter(
                                        Q(materia__asignatura__nombre__icontains=fullname) | Q(
                                            materia__asignatura__nombre__icontains=shortname), status=True,
                                        materiaasignadaretiro__isnull=True, matricula=matricula)[0]
                                    matricula.inscripcion.actividadessakaialumno_set.filter(status=True,tipo=4,materia=materiaasignada.materia,
                                                                                            inscripcion=matricula.inscripcion,
                                                                                            materia__nivel__periodo=periodo).delete()
                                    if not ActividadesSakaiAlumno.objects.filter(tipo=4, status=True,
                                                                                 inscripcion=matricula.inscripcion,
                                                                                 materia=materiaasignada.materia).exists():
                                        actividadguardada = ActividadesSakaiAlumno(inscripcion=matricula.inscripcion,
                                                                                   materia=materiaasignada.materia,
                                                                                   nombreactividadsakai=fullname,
                                                                                   tipo=4, nota=nota,
                                                                                   notaposible=40
                                                                                   )
                                        actividadguardada.save()
                                    else:
                                        actividadguardada = ActividadesSakaiAlumno.objects.filter(tipo=4, status=True,
                                                                                                  inscripcion=matricula.inscripcion,
                                                                                                  materia=materiaasignada.materia)[0]
                                        if nota > actividadguardada.nota:
                                            actividadguardada.nota = nota
                                            actividadguardada.notaposible = 40
                                            actividadguardada.save()

                log(u'Migración de datos de examenes : %s' % personasesion, request, "add")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'migrar_examenes_moodle_alumno':
            try:
                if 'inscripcion' in request.POST:
                    inscripcion = int(request.POST['inscripcion'])
                    inscripcion = Inscripcion.objects.get(id=inscripcion)
                matriculas = Matricula.objects.filter(status=True, nivel__periodo=periodo, inscripcion=inscripcion)
                for matricula in matriculas:
                    lista=""
                    # sql_examenes = " select DISTINCT user.username, curso.id,curso.fullname ,curso.shortname ,grade.finalgrade from mooc_grade_grades grade inner join mooc_user user on user.id=grade.userid inner join mooc_grade_items item on (item.id=grade.itemid) inner join mooc_course curso on (curso.id=item.courseid) where user.username ='" + str(matricula.inscripcion.persona.identificacion()) + "' order by username, id"
                    materiaasignada = matricula.materiaasignada_set.filter(status=True, materiaasignadaretiro__isnull=True).order_by('materia__id').distinct()
                    for x in materiaasignada:
                        if x.materia.idcursomoodle != materiaasignada.order_by('-materia__id')[0].materia.idcursomoodle:
                            if x.materia.idcursomoodle:
                                lista += str(x.materia.idcursomoodle) + ","
                        else:
                            lista += str(x.materia.idcursomoodle)
                    cursor = connections['db_moodle_virtual'].cursor()
                    sql_examenes="""
                        SELECT DISTINCT co.fullname, co.shortname, notatest.grade
                        FROM mooc_quiz_grades notatest
                        INNER JOIN mooc_quiz test ON notatest.quiz=test.id
                        INNER JOIN mooc_user u ON u.id=notatest.userid
                        INNER JOIN mooc_course co ON  co.id=test.course
                        WHERE notatest.grade >= 0 AND u.idnumber='%s' AND not co.id IN (%s)
                    """ %(str(matricula.inscripcion.persona.identificacion()), lista)
                    cursor.execute(sql_examenes)
                    datosexamenes = cursor.fetchall()
                    for dato in datosexamenes:
                        nota = dato[2] if dato else 0
                        fullname = dato[0] if dato else ""
                        shortname = dato[1] if dato else ""
                        if nota:
                            if MateriaAsignada.objects.filter(Q(materia__asignatura__nombre__icontains=fullname) | Q(materia__asignatura__nombre__icontains=shortname),
                                                              status=True, materiaasignadaretiro__isnull=True,matricula=matricula).exists():
                                materiaasignada =  MateriaAsignada.objects.filter(Q(materia__asignatura__nombre__icontains=fullname) | Q(materia__asignatura__nombre__icontains=shortname), status=True, materiaasignadaretiro__isnull=True, matricula=matricula)[0]

                                if not ActividadesSakaiAlumno.objects.filter(tipo=4, status=True, inscripcion=inscripcion,
                                                                             materia=materiaasignada.materia).exists():
                                    actividadguardada = ActividadesSakaiAlumno(inscripcion=inscripcion,
                                                                               materia=materiaasignada.materia,
                                                                               nombreactividadsakai=fullname,
                                                                               tipo=4, nota=nota,
                                                                               notaposible=40
                                                                               )
                                    actividadguardada.save()
                                else:
                                    actividadguardada = ActividadesSakaiAlumno.objects.filter(tipo=4, status=True,
                                                                                              inscripcion=inscripcion,
                                                                                              materia=materiaasignada.materia)[0]
                                    if nota > actividadguardada.nota:
                                        actividadguardada.nota = nota
                                        actividadguardada.notaposible = 40
                                        actividadguardada.save()
                log(u'Migración de datos de examenes moodle de un alumno: %s %s' % (personasesion, inscripcion), request, "add")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'pasar_notas_finales_alumno':
            try:
                # inscripcion = Inscripcion.objects.get(pk=request.POST['id'], status=True)
                # for materiaasignada in inscripcion.materias(periodo):
                #     ex1=0
                #     n1=0
                #     campo = materiaasignada.campo("N1")
                #     n1=inscripcion.promedio_general_por_asignatura_migradas(materiaasignada.materia.id)
                #     if n1==0:
                #         n1=0.1
                #     if n1>=0:
                #         actualizar_nota_planificacion(materiaasignada.id, "N1", n1)
                #         auditorianotas = AuditoriaNotas(evaluaciongenerica=campo, manual=False, calificacion=n1)
                #         auditorianotas.save(request)
                #     if materiaasignada.materia.esintroductoria:
                #         ex1=30
                #     else:
                #         ex1 = inscripcion.obtener_nota_examen_admision_virtual(materiaasignada.materia.id)
                #     if ex1 >=0:
                #         campo = materiaasignada.campo("EX1")
                #         actualizar_nota_planificacion(materiaasignada.id, "EX1", ex1)
                #         auditorianotas = AuditoriaNotas(evaluaciongenerica=campo, manual=False, calificacion=ex1)
                #         auditorianotas.save(request)
                # # inscripcion.matricula().promedionotasvirtual = inscripcion.matricula().inscripcion.total_estudiante_sobre_ponderacion_migrada(periodo)
                # # inscripcion.matricula().save()
                # inscripcion.matricula_admision_virtual(periodo).promedionotasvirtual = inscripcion.matricula_admision_virtual(periodo).inscripcion.total_estudiante_sobre_ponderacion_migrada(periodo)
                # inscripcion.matricula_admision_virtual(periodo).save()
                #
                inscripcion = Inscripcion.objects.get(pk=request.POST['id'], status=True)
                for materiaasignada in inscripcion.materias(periodo):
                    for notasmooc in materiaasignada.materia.notas_de_moodle(inscripcion.persona):
                        campo = materiaasignada.campo(notasmooc[1].upper())
                        if campo:
                            if type(notasmooc[0]) is Decimal:
                                if null_to_decimal(campo.valor) != float(notasmooc[0]):
                                    actualizar_nota_planificacion(materiaasignada.id, notasmooc[1].upper(), notasmooc[0])
                                    auditorianotas = AuditoriaNotas(evaluaciongenerica=campo, manual=False,
                                                                    calificacion=notasmooc[0])
                                    auditorianotas.save(request)
                            else:
                                if null_to_decimal(campo.valor) != float(0):
                                    actualizar_nota_planificacion(materiaasignada.id, notasmooc[1].upper(), notasmooc[0])
                                    auditorianotas = AuditoriaNotas(evaluaciongenerica=campo, manual=False, calificacion=0)
                                    auditorianotas.save(request)
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al actualizar los datos."})

        elif action == 'actualizar_record_academico_estudiante':
            try:
                inscripcion = Inscripcion.objects.get(pk=request.POST['id'], status=True)
                for materiaasignada in inscripcion.materias(periodo):
                    materiaasignada.materia.cerrado=False
                    materiaasignada.materia.fechacierre = datetime.now().date()
                    materiaasignada.materia.save()

                    materiaasignada.cerrado = True
                    materiaasignada.save()
                    materiaasignada.actualiza_estado()
                    materiaasignada.cierre_materia_asignada()

                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al actualizar los datos."})

        elif action == 'eliminar_matricula':
            try:
                inscripcion = Inscripcion.objects.get(pk=request.POST['id'], status=True)
                matricula = Matricula.objects.filter(inscripcion=inscripcion,nivel__periodo=periodo)[0]
                for materiaasignada in MateriaAsignada.objects.filter(matricula=matricula):
                    materiaasignada.delete()
                matricula.delete()
                # for historico in inscripcion.historicorecordacademico_set.all():
                #     historico.delete()
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al actualizar los datos."})

        elif action == 'eliminar_inscripcion':
            try:
                inscripcion = Inscripcion.objects.get(pk=request.POST['id'], status=True)
                inscripcion.delete()
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al actualizar los datos."})

        elif action == 'pasar_notas_finales':
            try:
                if 'carr' in request.POST:
                    idcarrera = int(request.POST['carr'])
                    carrera = Carrera.objects.get(id=idcarrera)
                matriculas = Matricula.objects.filter(status=True, nivel__periodo=periodo, inscripcion__carrera=carrera).distinct().order_by('inscripcion__persona__apellido1', 'inscripcion__persona__apellido2', 'inscripcion__persona__nombres')
                # for matricula in matriculas:
                #     for materiaasignada in matricula.inscripcion.materias(periodo):
                #         ex1 = 0
                #         n1 = 0
                #         campo = materiaasignada.campo("N1")
                #         n1 = matricula.inscripcion.promedio_general_por_asignatura_migradas(materiaasignada.materia.id)
                #         if n1 == 0:
                #             n1 = 0.1
                #         if n1>=0:
                #             actualizar_nota_planificacion(materiaasignada.id, "N1", n1)
                #             auditorianotas = AuditoriaNotas(evaluaciongenerica=campo, manual=False, calificacion=n1)
                #             auditorianotas.save(request)
                #         if materiaasignada.materia.esintroductoria:
                #             ex1 = 30
                #         else:
                #             ex1 = matricula.inscripcion.obtener_nota_examen_admision_virtual(materiaasignada.materia.id)
                #         if ex1 >=0:
                #             campo = materiaasignada.campo("EX1")
                #             actualizar_nota_planificacion(materiaasignada.id, "EX1", ex1)
                #             auditorianotas = AuditoriaNotas(evaluaciongenerica=campo, manual=False, calificacion=ex1)
                #             auditorianotas.save(request)
                #     matricula.promedionotasvirtual=matricula.inscripcion.total_estudiante_sobre_ponderacion_migrada(periodo)
                #     matricula.save()
                for matricula in matriculas:
                    for materiaasignada in matricula.inscripcion.materias(periodo):
                        for notasmooc in materiaasignada.materia.notas_de_moodle(matricula.inscripcion.persona):
                            campo = materiaasignada.campo(notasmooc[1].upper())
                            if type(notasmooc[0]) is Decimal:
                                if null_to_decimal(campo.valor) != float(notasmooc[0]):
                                    actualizar_nota_planificacion(materiaasignada.id, notasmooc[1].upper(), notasmooc[0])
                                    auditorianotas = AuditoriaNotas(evaluaciongenerica=campo, manual=False,
                                                                    calificacion=notasmooc[0])
                                    auditorianotas.save(request)
                            else:
                                if null_to_decimal(campo.valor) != float(0):
                                    actualizar_nota_planificacion(materiaasignada.id, notasmooc[1].upper(), notasmooc[0])
                                    auditorianotas = AuditoriaNotas(evaluaciongenerica=campo, manual=False, calificacion=0)
                                    auditorianotas.save(request)
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al actualizar los datos."})

        elif action == 'ver_horario':
            try:
                data = {}
                resultados = None
                data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.POST['id'])
                try:
                    sql = "select * from horario h where h.cedula='" + str(inscripcion.persona.identificacion()) + "'"
                    resultados = querymysqlconsulta(sql, True)
                except Exception as ex:
                    pass
                data['resultados'] = resultados
                template = get_template("inscripciones_admision/ver_horario.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'data': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'ver_errores_foros':
            materia=None
            try:
                __author__ = 'Unemi'
                carrera = None
                if 'carr' in request.POST:
                    idcarrera = int(request.POST['carr'])
                    carrera = Carrera.objects.get(id=idcarrera)
                matriculas = Matricula.objects.filter(status=True, nivel__periodo=periodo,
                                                      inscripcion__modalidad__id=3,
                                                      inscripcion__carrera=carrera).distinct().order_by(
                    'inscripcion__persona__apellido1', 'inscripcion__persona__apellido2',
                    'inscripcion__persona__nombres')

                # matriculas = Matricula.objects.filter(status=True, inscripcion_id=52715)
                title = easyxf(
                    'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                font_style = XFStyle()
                font_style.font.bold = True
                font_style2 = XFStyle()
                font_style2.font.bold = False
                wb = Workbook(encoding='utf-8')
                ws = wb.add_sheet('errores')
                ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                response = HttpResponse(content_type="application/ms-excel")
                response['Content-Disposition'] = 'attachment; filename=errores' + random.randint(1,
                                                                                                  10000).__str__() + '.xls'
                columns = [
                    (u"caso .", 5000),
                    (u"asignatura .", 5000),
                    (u"titulo_topico .", 5000),
                    (u"asigname_topico .", 5000),
                    (u"titulo_foro .", 5000),
                    (u"asigname_foro .", 5000),
                ]
                row_num = 4
                for col_num in range(len(columns)):
                    ws.write(row_num, col_num, columns[col_num][0], font_style)
                    ws.col(col_num).width = columns[col_num][1]
                date_format = xlwt.XFStyle()
                date_format.num_format_str = 'yyyy/mm/dd'
                row_num = 5
                for matricula in matriculas:
                    for materiaasignada in matricula.materiaasignada_set.filter(status=True):
                        materia = materiaasignada.materia.nombre_completo()
                        for foro in matricula.inscripcion.lista_foros_calificados_asignatura(
                                materiaasignada.materia.codigosakai):
                            id_foro = foro[0] if foro[1] else ""
                            titulo_topico = foro[1].strip() if foro[1] else ""
                            asigname_topico = foro[2].strip() if foro[2] else ""
                            titulo_foro = foro[3].strip() if foro[3] else ""
                            asigname_foro = foro[4].strip() if foro[4] else ""
                            try:
                                calificaciones = matricula.inscripcion.calificacion_foro(
                                    materiaasignada.materia.codigosakai, asigname_topico, titulo_topico)
                            except Exception as ex:
                                calificaciones = None
                            if not calificaciones:
                                try:
                                    titulo_foro = titulo_foro.replace(u'.', '')
                                    calificaciones = matricula.inscripcion.calificacion_foro(
                                        materiaasignada.materia.codigosakai, asigname_foro, titulo_foro)
                                except Exception as ex:
                                    calificaciones = None
                            if calificaciones.__len__()<=0:
                                ws.write(row_num, 0, matricula.inscripcion.persona.nombre_completo_inverso(), font_style2)
                                ws.write(row_num, 1, materiaasignada.materia.nombre_completo(), font_style2)
                                ws.write(row_num, 2, titulo_topico, font_style2)
                                ws.write(row_num, 3, asigname_topico, font_style2)
                                ws.write(row_num, 4, titulo_foro, font_style2)
                                ws.write(row_num, 5, asigname_foro, font_style2)
                                row_num += 1
                wb.save(response)
                return response
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos. %s %s" % (ex,materia)})

        elif action == 'listado_alumnos_promedios':
            campo1=None
            try:
                __author__ = 'Unemi'
                carrera = None
                if 'carr' in request.POST:
                    idcarrera = int(request.POST['carr'])
                    carrera = Carrera.objects.get(id=idcarrera)
                matriculas = Matricula.objects.filter(status=True, nivel__periodo=periodo, inscripcion__modalidad__id=3,
                                                      inscripcion__carrera=carrera).distinct().order_by(
                    'inscripcion__persona__apellido1', 'inscripcion__persona__apellido2',
                    'inscripcion__persona__nombres')
                style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
                style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                style1 = easyxf(num_format_str='D-MMM-YY')
                font_style = XFStyle()
                font_style.font.bold = True
                font_style2 = XFStyle()
                font_style2.font.bold = False
                wb = Workbook(encoding='utf-8')
                ws = wb.add_sheet('alumnos_online_promedio')
                ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                response = HttpResponse(content_type="application/ms-excel")
                response['Content-Disposition'] = 'attachment; filename=alumnos_online_promedio de '+carrera.nombre_completo() +" " + random.randint(1, 10000).__str__() + '.xls'
                columns = [
                    (u"Cedula .", 5000),
                    (u"Carrera .", 5000),
                    (u"Apellidos .", 5000),
                    (u"Nombres .", 5000),
                    (u"Asignatura", 5000),
                    (u"Paralelo", 5000),
                    (u"Total gestion", 5000),
                    # (u"Examen", 5000),
                    # (u"gestion y examen", 5000),
                    # (u"estado", 5000),
                    # (u"Gestion ponderacion ", 5000),
                    # (u"Examen ponderacion ", 5000),
                    # (u"Ponderacion", 5000),
                    # (u"Total", 5000),
                    # (u"cantidad examen rendidos", 5000),
                    # (u"cantidad materia aprobadas", 5000),
                    # (u"cantidad materia reprobada", 5000),
                ]
                row_num = 4
                for col_num in range(len(columns)):
                    ws.write(row_num, col_num, columns[col_num][0], font_style)
                    ws.col(col_num).width = columns[col_num][1]
                date_format = xlwt.XFStyle()
                date_format.num_format_str = 'yyyy/mm/dd'
                row_num = 5
                for matricula in matriculas:
                    materiasasignadas = matricula.materiaasignada_set.filter(status=True, materiaasignadaretiro__isnull=True, materia__nivel__periodo=periodo, materia__esintroductoria=False).distinct()
                    for materiaasignada in materiasasignadas:
                        gestionyexamen=0
                        campo1 = str(matricula.inscripcion.persona.identificacion())
                        campo2 = str(matricula.inscripcion.carrera.nombre_completo())
                        campo3 = str( matricula.inscripcion.persona.apellido1 + " " + matricula.inscripcion.persona.apellido2)
                        campo4 = str(matricula.inscripcion.persona.nombres)
                        # totalfinal = matricula.inscripcion.total_estudiante_sobre_ponderacion_migrada(periodo)
                        estado =""
                        materia = materiaasignada.materia
                        paralelo = materiaasignada.materia.paralelo
                        try:
                            gestion_sobre_70 = matricula.inscripcion.promedio_general_por_asignatura_migradas1(materiaasignada.materia.id)
                        except Exception as ex:
                            gestion_sobre_70 = 0

                        # try:
                        #     gestionponderacion = matricula.inscripcion.gestion_ponderacion_por_asignatura_migradas(materiaasignada.materia.id)
                        # except Exception as ex:
                        #     gestionponderacion = 0
                        #
                        # try:
                        #     ponderacion = matricula.inscripcion.porcentaje_equivalente_asignatura( materiaasignada.materia.id)
                        # except Exception as ex:
                        #     ponderacion = 0
                        #
                        # try:
                        #     examen = matricula.inscripcion.nota_examen_migrada(materiaasignada.materia.id)
                        # except Exception as ex:
                        #     examen = 0
                        #
                        # try:
                        #     examenponderacion = matricula.inscripcion.examen_ponderacion_por_asignatura_migradas( materiaasignada.materia.id)
                        # except Exception as ex:
                        #     examenponderacion = 0
                        #
                        # try:
                        #     total = matricula.inscripcion.porcentaje_por_asignatura_con_examen_migradas( materiaasignada.materia.id)
                        # except Exception as ex:
                        #     total = 0
                        #
                        # gestionyexamen = null_to_decimal(gestion_sobre_70 + (examen if examen else 0),2)
                        # if gestionyexamen >=70:
                        #     estado ="A"
                        # else:
                        #     estado="N"
                        #
                        # cantidad_examen_rendidos = matricula.cantidad_examen_rindio_admision_virtual()
                        # cantidad_materia_aprobadas = matricula.cantidad_materias_aprobadas(periodo)
                        # cantidad_materia_reprobada = matricula.cantidad_materias_reprobadas(periodo)
                        ws.write(row_num, 0, campo1, font_style2)
                        ws.write(row_num, 1, campo2, font_style2)
                        ws.write(row_num, 2, campo3, font_style2)
                        ws.write(row_num, 3, campo4, font_style2)
                        ws.write(row_num, 4, materia.nombre_completo(), font_style2)
                        ws.write(row_num, 5, paralelo, font_style2)
                        ws.write(row_num, 6, gestion_sobre_70 if gestion_sobre_70 else 0, font_style2)
                        # ws.write(row_num, 7, examen if examen else 0, font_style2)
                        # ws.write(row_num, 8, gestionyexamen if gestionyexamen else 0, font_style2)
                        # ws.write(row_num, 9, estado, font_style2)
                        # ws.write(row_num, 10, gestionponderacion if gestionponderacion else 0, font_style2)
                        # ws.write(row_num, 11, examenponderacion if examenponderacion else 0, font_style2)
                        # ws.write(row_num, 12, ponderacion if ponderacion else 0, font_style2)
                        # ws.write(row_num, 13, total if total else 0, font_style2)
                        # ws.write(row_num, 14, totalfinal if totalfinal else 0, font_style2)
                        # ws.write(row_num, 15, cantidad_examen_rendidos if cantidad_examen_rendidos else 0, font_style2)
                        # ws.write(row_num, 16, cantidad_materia_aprobadas if cantidad_materia_aprobadas else 0, font_style2)
                        # ws.write(row_num, 17, cantidad_materia_reprobada if cantidad_materia_reprobada else 0, font_style2)
                        row_num += 1
                    row_num += 1
                wb.save(response)
                return response
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos. %s %s" % (ex,campo1)})

        elif action == 'listado_alumnos_promedios_columna':
            campo1=None
            try:
                __author__ = 'Unemi'
                carrera = None
                if 'carr' in request.POST:
                    idcarrera = int(request.POST['carr'])
                    carrera = Carrera.objects.get(id=idcarrera)
                matriculas = Matricula.objects.filter(status=True, nivel__periodo=periodo, inscripcion__modalidad__id=3,
                                                      inscripcion__carrera=carrera).exclude(
                    inscripcion__persona__ppl=True, inscripcion__persona__pais_id=8).distinct().order_by(
                    'inscripcion__persona__apellido1', 'inscripcion__persona__apellido2',
                    'inscripcion__persona__nombres')
                style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
                style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                title = easyxf(
                    'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                style1 = easyxf(num_format_str='D-MMM-YY')
                font_style = XFStyle()
                font_style.font.bold = True
                font_style2 = XFStyle()
                font_style2.font.bold = False
                wb = Workbook(encoding='utf-8')
                ws = wb.add_sheet('alumnos_online_promedio')
                ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                response = HttpResponse(content_type="application/ms-excel")
                response['Content-Disposition'] = 'attachment; filename=alumnos_online_promedio de '+carrera.nombre_completo() +" " + random.randint(1, 10000).__str__() + '.xls'
                columns = [
                    (u"Cedula .", 5000),
                    (u"Carrera .", 5000),
                    (u"Apellidos .", 5000),
                    (u"Nombres .", 5000),
                    (u"Institucional .", 5000),
                    (u"Username .", 5000),
                    (u"Cedula .", 5000),
                    (u"Correo .", 5000),
                    (u"Telefono .", 5000),
                    (u"ppl .", 5000),
                    (u"residencia .", 5000),
                    (u"discapacidad .", 5000),
                    (u"Paralelo", 5000),
                    (u"Total Final", 5000),
                    (u"cantidad examen rendidos", 5000),
                    (u"cantidad materia aprobadas", 5000),
                    (u"cantidad materia reprobada", 5000),

                    (u"Asignatura 1", 5000),
                    (u"Gestion sobre 70 Asig 1", 5000),
                    (u"Examen Asig 1", 5000),
                    (u"gestion y examen Asig 1", 5000),
                    (u"estado Asig 1", 5000),
                    (u"Gestion ponderacion Asig 1", 5000),
                    (u"Examen ponderacion Asig 1", 5000),
                    (u"Ponderacion Asig 1", 5000),
                    (u"Total Asig 1", 5000),

                    (u"Asignatura 2", 5000),
                    (u"Gestion sobre 70 Asig 2", 5000),
                    (u"Examen Asig 2", 5000),
                    (u"gestion y examen Asig 2", 5000),
                    (u"estado Asig 2", 5000),
                    (u"Gestion ponderacion Asig 2", 5000),
                    (u"Examen ponderacion Asig 2", 5000),
                    (u"Ponderacion Asig 2", 5000),
                    (u"Total Asig 2", 5000),

                    (u"Asignatura 3", 5000),
                    (u"Gestion sobre 70 Asig 3", 5000),
                    (u"Examen Asig 3", 5000),
                    (u"gestion y examen Asig 3", 5000),
                    (u"estado Asig 3", 5000),
                    (u"Gestion ponderacion Asig 3", 5000),
                    (u"Examen ponderacion Asig 3", 5000),
                    (u"Ponderacion Asig 3", 5000),
                    (u"Total Asig 3", 5000),

                    (u"Asignatura 4", 5000),
                    (u"Gestion sobre 70 Asig 4", 5000),
                    (u"Examen Asig 4", 5000),
                    (u"gestion y examen Asig 4", 5000),
                    (u"estado Asig 4", 5000),
                    (u"Gestion ponderacion Asig 4", 5000),
                    (u"Examen ponderacion Asig 4", 5000),
                    (u"Ponderacion Asig 4", 5000),
                    (u"Total Asig 4", 5000),

                    (u"Asignatura 5", 5000),
                    (u"Gestion sobre 70 Asig 5", 5000),
                    (u"Examen Asig 5", 5000),
                    (u"gestion y examen Asig 5", 5000),
                    (u"estado Asig 5", 5000),
                    (u"Gestion ponderacion Asig 5", 5000),
                    (u"Examen ponderacion Asig 5", 5000),
                    (u"Ponderacion Asig 5", 5000),
                    (u"Total Asig 5", 5000),
                ]
                row_num = 4
                for col_num in range(len(columns)):
                    ws.write(row_num, col_num, columns[col_num][0], font_style)
                    ws.col(col_num).width = columns[col_num][1]
                date_format = xlwt.XFStyle()
                date_format.num_format_str = 'yyyy/mm/dd'
                row_num = 5
                for matricula in matriculas:
                    campo1 = str(matricula.inscripcion.persona.identificacion())
                    campo2 = str(matricula.inscripcion.carrera.nombre_completo())
                    campo3 = str( matricula.inscripcion.persona.apellido1 + " " + matricula.inscripcion.persona.apellido2)
                    campo4 = str(matricula.inscripcion.persona.nombres)
                    usuario = matricula.inscripcion.persona.usuario.username if not matricula.inscripcion.persona.usuario == None else "NO"
                    institucional = matricula.inscripcion.persona.emailinst if matricula.inscripcion.persona.emailinst else "NO"
                    correo = matricula.inscripcion.persona.email if matricula.inscripcion.persona.email else "NO"
                    telefono = matricula.inscripcion.persona.telefono if matricula.inscripcion.persona.telefono else "NO"
                    discapacidad = "SI" if matricula.inscripcion.persona.perfilinscripcion_set.filter(tienediscapacidad=True, status=True).exists() else "NO"
                    ppl = "SI" if matricula.inscripcion.persona.ppl else "NO"
                    residencia = str(matricula.inscripcion.persona.pais.nombre) if matricula.inscripcion.persona.pais else ""
                    cantidad_examen_rendidos = matricula.cantidad_examen_rindio_admision_virtual()
                    cantidad_materia_aprobadas = matricula.cantidad_materias_aprobadas(periodo)
                    cantidad_materia_reprobada = matricula.cantidad_materias_reprobadas(periodo)
                    totalfinal = matricula.inscripcion.total_estudiante_sobre_ponderacion_migrada(periodo)
                    ws.write(row_num, 0, campo1, font_style2)
                    ws.write(row_num, 1, campo2, font_style2)
                    ws.write(row_num, 2, campo3, font_style2)
                    ws.write(row_num, 3, campo4, font_style2)
                    ws.write(row_num, 4, institucional, font_style2)
                    ws.write(row_num, 5, usuario, font_style2)
                    ws.write(row_num, 6, campo1, font_style2)
                    ws.write(row_num, 7, correo, font_style2)
                    ws.write(row_num, 8, telefono, font_style2)
                    ws.write(row_num, 9, ppl, font_style2)
                    ws.write(row_num, 10, residencia, font_style2)
                    ws.write(row_num, 11, discapacidad, font_style2)
                    ws.write(row_num, 13, totalfinal, font_style2)
                    ws.write(row_num, 14, cantidad_examen_rendidos if cantidad_examen_rendidos else 0, font_style2)
                    ws.write(row_num, 15, cantidad_materia_aprobadas if cantidad_materia_aprobadas else 0, font_style2)
                    ws.write(row_num, 16, cantidad_materia_reprobada if cantidad_materia_reprobada else 0, font_style2)
                    materiasasignadas = matricula.mis_materias_sin_retiro()
                    col_num =16
                    for materiaasignada in materiasasignadas:
                        estado =""
                        materia = materiaasignada.materia
                        paralelo = materiaasignada.materia.paralelo
                        try:
                            gestion_sobre_70 = matricula.inscripcion.promedio_general_por_asignatura_migradas(materiaasignada.materia.id)
                        except Exception as ex:
                            gestion_sobre_70 = 0

                        try:
                            gestionponderacion = matricula.inscripcion.gestion_ponderacion_por_asignatura_migradas(materiaasignada.materia.id)
                        except Exception as ex:
                            gestionponderacion = 0

                        try:
                            ponderacion = matricula.inscripcion.porcentaje_equivalente_asignatura( materiaasignada.materia.id)
                        except Exception as ex:
                            ponderacion = 0

                        try:
                            examen = matricula.inscripcion.nota_examen_migrada(materiaasignada.materia.id)
                        except Exception as ex:
                            examen = 0

                        try:
                            examenponderacion = matricula.inscripcion.examen_ponderacion_por_asignatura_migradas( materiaasignada.materia.id)
                        except Exception as ex:
                            examenponderacion = 0

                        try:
                            total = matricula.inscripcion.porcentaje_por_asignatura_con_examen_migradas( materiaasignada.materia.id)
                        except Exception as ex:
                            total = 0

                        gestionyexamen = null_to_decimal(gestion_sobre_70 + (examen if examen else 0),2)
                        if gestionyexamen >=70:
                            estado ="A"
                        else:
                            estado="N"
                        if col_num==16:
                            ws.write(row_num, 12, paralelo, font_style2)
                        col_num += 1
                        ws.write(row_num, col_num, materia.nombre_completo(), font_style2)
                        col_num += 1
                        ws.write(row_num, col_num, gestion_sobre_70, font_style2)
                        col_num += 1
                        ws.write(row_num, col_num, examen, font_style2)
                        col_num += 1
                        ws.write(row_num, col_num, gestionyexamen, font_style2)
                        col_num += 1
                        ws.write(row_num, col_num, estado, font_style2)
                        col_num += 1
                        ws.write(row_num, col_num, gestionponderacion, font_style2)
                        col_num += 1
                        ws.write(row_num, col_num, examenponderacion, font_style2)
                        col_num += 1
                        ws.write(row_num, col_num, ponderacion, font_style2)
                        col_num += 1
                        ws.write(row_num, col_num, total, font_style2)
                    row_num += 1
                wb.save(response)
                return response
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos. %s %s" % (ex,campo1)})

        elif action == 'ver_errores_tareas':
            try:
                __author__ = 'Unemi'
                carrera = None
                if 'carr' in request.POST:
                    idcarrera = int(request.POST['carr'])
                    carrera = Carrera.objects.get(id=idcarrera)
                matriculas = Matricula.objects.filter(status=True, nivel__periodo=periodo,
                                                      inscripcion__modalidad__id=3,
                                                      inscripcion__carrera=carrera).distinct().order_by(
                    'inscripcion__persona__apellido1', 'inscripcion__persona__apellido2',
                    'inscripcion__persona__nombres')

                # matriculas = Matricula.objects.filter(status=True, inscripcion_id=52715)
                title = easyxf( 'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                font_style = XFStyle()
                font_style.font.bold = True
                font_style2 = XFStyle()
                font_style2.font.bold = False
                wb = Workbook(encoding='utf-8')
                ws = wb.add_sheet('errores_tarea')
                ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                response = HttpResponse(content_type="application/ms-excel")
                response['Content-Disposition'] = 'attachment; filename=errores_tarea' + random.randint(1,
                                                                                                  10000).__str__() + '.xls'
                columns = [
                    (u"caso .", 5000),
                    (u"asignatura .", 5000),
                    (u"tarea .", 5000),
                ]
                row_num = 4
                for col_num in range(len(columns)):
                    ws.write(row_num, col_num, columns[col_num][0], font_style)
                    ws.col(col_num).width = columns[col_num][1]
                date_format = xlwt.XFStyle()
                date_format.num_format_str = 'yyyy/mm/dd'
                row_num = 5
                for matricula in matriculas:
                    for materiaasignada in matricula.materiaasignada_set.filter(status=True):
                        tareas = matricula.inscripcion.lista_tareas_por_asignatura(materiaasignada.materia.codigosakai)
                        actividadguardada = None
                        for tarea in tareas:
                            id_tarea = tarea[0]
                            id_xml_tarea = tarea[1]
                            xmltext = minidom.parseString(id_xml_tarea)
                            itemlist = xmltext.getElementsByTagName("assignment")
                            nombre_tarea = (itemlist[0].getAttribute('title')).upper()
                            resultados_enviotarea = matricula.inscripcion.envio_tarea(id_tarea, materiaasignada.materia.codigosakai)
                            scaled_grade = 0
                            puntos_posibles = 0
                            if resultados_enviotarea:
                                id_xml_tarea = resultados_enviotarea[0][6]
                                xmltext = minidom.parseString(id_xml_tarea)
                                itemlist = xmltext.getElementsByTagName("submission")
                                if itemlist[0].getAttribute('scaled_grade') != '':
                                    try:
                                        scaled_grade = null_to_decimal(int(itemlist[0].getAttribute('scaled_grade')) / 100)
                                    except Exception as ex:
                                        scaled_grade = None
                            try:
                                puntos_posibles = matricula.inscripcion.puntos_posibles_por_tarea(id_tarea, materiaasignada.materia.codigosakai)
                            except Exception as ex:
                                puntos_posibles = 0

                            if puntos_posibles.__len__() > 0 and not puntos_posibles == 0:
                                if scaled_grade == None:
                                    ws.write(row_num, 0, matricula.inscripcion.persona.nombre_completo_inverso(), font_style2)
                                    ws.write(row_num, 1, materiaasignada.materia.nombre_completo(), font_style2)
                                    ws.write(row_num, 2, nombre_tarea, font_style2)
                                    row_num += 1
                wb.save(response)
                return response
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos. %s" % ex})

        elif action == 'ver_errores_test':
            try:
                __author__ = 'Unemi'
                carrera = None
                if 'carr' in request.POST:
                    idcarrera = int(request.POST['carr'])
                    carrera = Carrera.objects.get(id=idcarrera)
                matriculas = Matricula.objects.filter(status=True, nivel__periodo=periodo,
                                                      inscripcion__modalidad__id=3,
                                                      inscripcion__carrera=carrera).distinct().order_by(
                    'inscripcion__persona__apellido1', 'inscripcion__persona__apellido2',
                    'inscripcion__persona__nombres')

                # matriculas = Matricula.objects.filter(status=True, inscripcion_id=52715)
                title = easyxf(
                    'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                font_style = XFStyle()
                font_style.font.bold = True
                font_style2 = XFStyle()
                font_style2.font.bold = False
                wb = Workbook(encoding='utf-8')
                ws = wb.add_sheet('errores_test')
                ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                response = HttpResponse(content_type="application/ms-excel")
                response['Content-Disposition'] = 'attachment; filename=errores_test' + random.randint(1,
                                                                                                  10000).__str__() + '.xls'
                columns = [
                    (u"caso .", 5000),
                    (u"asignatura .", 5000),
                    (u"test .", 5000),
                ]
                row_num = 4
                for col_num in range(len(columns)):
                    ws.write(row_num, col_num, columns[col_num][0], font_style)
                    ws.col(col_num).width = columns[col_num][1]
                date_format = xlwt.XFStyle()
                date_format.num_format_str = 'yyyy/mm/dd'
                row_num = 5
                for matricula in matriculas:
                    for materiaasignada in matricula.materiaasignada_set.filter(status=True):
                        for test in matricula.inscripcion.lista_test_asignatura(materiaasignada.materia.codigosakai):
                            id_test = test[0]
                            nombre_test = test[1]
                            try:
                                nota = matricula.inscripcion.calificacion_test(materiaasignada.materia.codigosakai, test[1])
                            except Exception as ex:
                                nota = None
                            try:
                                notaposible = matricula.inscripcion.puntos_posibles_test(materiaasignada.materia.codigosakai,test[1])
                            except Exception as ex:
                                notaposible = None
                            if notaposible:
                                if nota == None:
                                    ws.write(row_num, 0, matricula.inscripcion.persona.nombre_completo_inverso(), font_style2)
                                    ws.write(row_num, 1, materiaasignada.materia.nombre_completo(), font_style2)
                                    ws.write(row_num, 2, nombre_test, font_style2)
                                    row_num += 1
                wb.save(response)
                return response
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos. %s" % ex})

        elif action == 'division_asignaturas':
            try:
                __author__ = 'Unemi'
                title = easyxf(
                    'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                font_style = XFStyle()
                font_style.font.bold = True
                font_style2 = XFStyle()
                font_style2.font.bold = False
                wb = Workbook(encoding='utf-8')
                ws = wb.add_sheet('errores_test')
                ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                response = HttpResponse(content_type="application/ms-excel")
                response['Content-Disposition'] = 'attachment; filename=cantidad_paralelos' + random.randint(1, 10000).__str__() + '.xls'
                columns = [
                    (u"Asignatura .", 5000),
                    (u"Paralelos .", 5000),
                ]
                row_num = 4
                for col_num in range(len(columns)):
                    ws.write(row_num, col_num, columns[col_num][0], font_style)
                    ws.col(col_num).width = columns[col_num][1]
                date_format = xlwt.XFStyle()
                date_format.num_format_str = 'yyyy/mm/dd'
                row_num = 5
                for asignatura in Asignatura.objects.filter(materia__nivel__periodo=periodo,materia__nivel__modalidad=3).distinct():
                    nombre=asignatura.nombre
                    cantidad=asignatura.materia_set.filter(nivel__periodo=periodo,modeloevaluativo_id=11).count()
                    ws.write(row_num, 0, nombre, font_style2)
                    ws.write(row_num, 1, cantidad, font_style2)
                    row_num += 1
                wb.save(response)
                return response
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos. %s" % ex})

        elif action == 'alumnos_matriculas_moodle':
            campo1=None
            try:
                __author__ = 'Unemi'
                matriculas = Matricula.objects.filter(status=True, nivel__periodo=periodo, nivel__sesion__id__in=[12,13]).distinct().order_by('inscripcion__persona__apellido1', 'inscripcion__persona__apellido2','inscripcion__persona__nombres')
                style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
                style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                style1 = easyxf(num_format_str='D-MMM-YY')
                font_style = XFStyle()
                font_style.font.bold = True
                font_style2 = XFStyle()
                font_style2.font.bold = False
                wb = Workbook(encoding='utf-8')
                ws = wb.add_sheet('alumnos_online_promedio')
                ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                response = HttpResponse(content_type="application/ms-excel")
                response['Content-Disposition'] = 'attachment; filename=listado_alumnos_asignatura_fila ' + random.randint(1, 10000).__str__() + '.xls'
                columns = [
                    (u"Cedula .", 5000),
                    (u"Carrera .", 5000),
                    (u"Apellidos .", 5000),
                    (u"Nombres .", 5000),
                    (u"Correo", 5000),
                    (u"Paralelo", 5000),
                    (u"ID Asignatura", 5000),
                    (u"Asignatura", 5000),
                    (u"Codigo Asignatura", 5000),
                ]
                row_num = 4
                for col_num in range(len(columns)):
                    ws.write(row_num, col_num, columns[col_num][0], font_style)
                    ws.col(col_num).width = columns[col_num][1]
                date_format = xlwt.XFStyle()
                date_format.num_format_str = 'yyyy/mm/dd'
                row_num = 5
                for matricula in matriculas:
                    materiasasignadas = matricula.inscripcion.materias(periodo)
                    for materiaasignada in materiasasignadas:
                        campo1 = str(matricula.inscripcion.persona.identificacion())
                        campo2 = str(matricula.inscripcion.carrera.nombre_completo())
                        campo3 = str( matricula.inscripcion.persona.apellido1 + " " + matricula.inscripcion.persona.apellido2)
                        campo4 = str(matricula.inscripcion.persona.nombres)
                        campo5 = str(matricula.inscripcion.persona.email)
                        paralelo = materiaasignada.materia.paralelo
                        materia = materiaasignada.materia
                        codigo = materia.asignatura.nombre[:2].upper()+"_"+paralelo + '_1S_2019'
                        ws.write(row_num, 0, campo1, font_style2)
                        ws.write(row_num, 1, campo2, font_style2)
                        ws.write(row_num, 2, campo3, font_style2)
                        ws.write(row_num, 3, campo4, font_style2)
                        ws.write(row_num, 4, campo5, font_style2)
                        ws.write(row_num, 5, paralelo, font_style2)
                        ws.write(row_num, 6, materia.id, font_style2)
                        ws.write(row_num, 7, materia.nombre_completo(), font_style2)
                        ws.write(row_num, 8, codigo, font_style2)
                        row_num += 1
                    # row_num += 1
                wb.save(response)
                return response
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos. %s %s" % (ex,campo1)})

        elif action == 'alumnos_matriculas_moodle_columna':
            campo1=None
            try:
                __author__ = 'Unemi'
                matriculas = Matricula.objects.filter(status=True, nivel__periodo=periodo,nivel__sesion__id__in=[12,13]).distinct().order_by('inscripcion__persona__apellido1', 'inscripcion__persona__apellido2', 'inscripcion__persona__nombres')
                style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
                style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                style1 = easyxf(num_format_str='D-MMM-YY')
                font_style = XFStyle()
                font_style.font.bold = True
                font_style2 = XFStyle()
                font_style2.font.bold = False
                wb = Workbook(encoding='utf-8')
                ws = wb.add_sheet('alumnos_online_promedio')
                ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                response = HttpResponse(content_type="application/ms-excel")
                response['Content-Disposition'] = 'attachment; filename=listado_alumnos_asignatura_columna ' + random.randint(1, 10000).__str__() + '.xls'
                columns = [
                    (u"Cedula", 5000),
                    (u"Sección", 5000),
                    (u"Carrera", 5000),
                    (u"Apellidos", 5000),
                    (u"Nombres", 5000),
                    (u"Usuario", 5000),
                    (u"Institucional", 5000),
                    (u"Correo Personal", 5000),
                    (u"Telefono", 5000),

                    (u"ID", 5000),
                    (u"Asignatura 1", 5000),
                    (u"Codigo PRE Virtual", 5000),
                    (u"Codigo presencial", 5000),
                    (u"Codigo Semestre", 5000),

                    (u"ID", 5000),
                    (u"Asignatura 2", 5000),
                    (u"Codigo PRE Virtual", 5000),
                    (u"Codigo presencial", 5000),
                    (u"Codigo Semestre", 5000),

                    (u"ID", 5000),
                    (u"Asignatura 3", 5000),
                    (u"Codigo PRE Virtual", 5000),
                    (u"Codigo presencial", 5000),
                    (u"Codigo Semestre", 5000),

                    (u"ID", 5000),
                    (u"Asignatura 4", 5000),
                    (u"Codigo PRE Virtual", 5000),
                    (u"Codigo presencial", 5000),
                    (u"Codigo Semestre", 5000),

                    (u"ID", 5000),
                    (u"Asignatura 5", 5000),
                    (u"Codigo PRE Virtual", 5000),
                    (u"Codigo presencial", 5000),
                    (u"Codigo Semestre", 5000),
                ]
                row_num = 4
                for col_num in range(len(columns)):
                    ws.write(row_num, col_num, columns[col_num][0], font_style)
                    ws.col(col_num).width = columns[col_num][1]
                date_format = xlwt.XFStyle()
                date_format.num_format_str = 'yyyy/mm/dd'
                row_num = 5
                for matricula in matriculas:
                    campo1 = str(matricula.inscripcion.persona.identificacion())
                    campo2 = matricula.nivel.sesion.nombre
                    campo3 = str(matricula.inscripcion.carrera.nombre_completo())
                    campo4 = str(matricula.inscripcion.persona.apellido1 + " " + matricula.inscripcion.persona.apellido2)
                    campo5 = str(matricula.inscripcion.persona.nombres)
                    campo6 = str(matricula.inscripcion.persona.usuario.username if not matricula.inscripcion.persona.usuario==None else "S/N")
                    campo7 = str(matricula.inscripcion.persona.emailinst)
                    campo8 = str(matricula.inscripcion.persona.email)
                    campo9 = str(matricula.inscripcion.persona.telefono)

                    ws.write(row_num, 0, campo1, font_style2)
                    ws.write(row_num, 1, campo2, font_style2)
                    ws.write(row_num, 2, campo3, font_style2)
                    ws.write(row_num, 3, campo4, font_style2)
                    ws.write(row_num, 4, campo5, font_style2)
                    ws.write(row_num, 5, campo6, font_style2)
                    ws.write(row_num, 6, campo7, font_style2)
                    ws.write(row_num, 7, campo8, font_style2)
                    ws.write(row_num, 8, campo9, font_style2)
                    materiasasignadas = matricula.materiaasignada_set.all()
                    col_num =8
                    for materiaasignada in materiasasignadas:
                        materia = materiaasignada.materia
                        paralelo = materiaasignada.materia.paralelo
                        codigo = materia.asignatura.nombre[:2].upper()+"_"+paralelo + '_1S_2019'
                        codigo_paralelo = paralelo + '_1S_2019_'+str(materiaasignada.materia.id)
                        col_num += 1
                        ws.write(row_num, col_num,  materiaasignada.materia.id, font_style2)
                        col_num += 1
                        ws.write(row_num, col_num, materia.asignatura.nombre, font_style2)
                        col_num += 1
                        ws.write(row_num, col_num, codigo, font_style2)
                        col_num += 1
                        ws.write(row_num, col_num, codigo_paralelo, font_style2)
                        col_num += 1
                        ws.write(row_num, col_num, codigo_paralelo, font_style2)
                    row_num += 1
                wb.save(response)
                return response
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos. %s %s" % (ex,campo1)})

        elif action == 'docentes_moodle_columna':
            campo1=None
            try:
                __author__ = 'Unemi'
                profesormaterias = ProfesorMateria.objects.filter(tipoprofesor__id=9,materia__nivel__periodo=periodo).distinct('profesor__persona__id')
                style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
                style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                style1 = easyxf(num_format_str='D-MMM-YY')
                font_style = XFStyle()
                font_style.font.bold = True
                font_style2 = XFStyle()
                font_style2.font.bold = False
                wb = Workbook(encoding='utf-8')
                ws = wb.add_sheet('alumnos_online_promedio')
                ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                response = HttpResponse(content_type="application/ms-excel")
                response['Content-Disposition'] = 'attachment; filename=distributivo_docentes ' + random.randint(1, 10000).__str__() + '.xls'
                columns = [
                    (u"Cedula", 5000),
                    (u"Sección", 5000),
                    (u"Apellidos", 5000),
                    (u"Nombres", 5000),
                    (u"Usuario", 5000),
                    (u"Institucional", 5000),
                    (u"Correo Personal", 5000),
                    (u"Telefono", 5000),
                ]
                row_num = 4
                for col_num in range(len(columns)):
                    ws.write(row_num, col_num, columns[col_num][0], font_style)
                    ws.col(col_num).width = columns[col_num][1]
                date_format = xlwt.XFStyle()
                date_format.num_format_str = 'yyyy/mm/dd'
                row_num = 5
                for profesores in profesormaterias:
                    campo1 = str(profesores.profesor.persona.identificacion())
                    campo2 = profesores.materia.nivel.sesion.nombre
                    campo3 = str( profesores.profesor.persona.apellido1 + " " +profesores.profesor.persona.apellido2)
                    campo4 = str(profesores.profesor.persona.nombres)
                    campo5 = str(profesores.profesor.persona.usuario.username if not profesores.profesor.persona.usuario == None else "S/N")
                    campo6 = str(profesores.profesor.persona.emailinst)
                    campo7 = str(profesores.profesor.persona.email)
                    campo8 = str(profesores.profesor.persona.telefono)

                    ws.write(row_num, 0, campo1, font_style2)
                    ws.write(row_num, 1, campo2, font_style2)
                    ws.write(row_num, 2, campo3, font_style2)
                    ws.write(row_num, 3, campo4, font_style2)
                    ws.write(row_num, 4, campo5, font_style2)
                    ws.write(row_num, 5, campo6, font_style2)
                    ws.write(row_num, 6, campo7, font_style2)
                    ws.write(row_num, 7, campo8, font_style2)
                    materiasasignadas = profesores.profesor.mis_materias_virtuales(periodo)
                    col_num = 7
                    for materiaasignada in materiasasignadas:
                        materia = materiaasignada.materia
                        paralelo = materiaasignada.materia.paralelo
                        codigo = paralelo + '_1S_2019_' + str(materiaasignada.materia.id)
                        col_num += 1
                        ws.write(row_num, col_num, materiaasignada.materia.asignaturamalla.malla.carrera.nombre, font_style2)
                        col_num += 1
                        ws.write(row_num, col_num, materiaasignada.materia.id, font_style2)
                        col_num += 1
                        ws.write(row_num, col_num, materia.asignatura.nombre, font_style2)
                        col_num += 1
                        ws.write(row_num, col_num, codigo, font_style2)
                        col_num += 1
                    row_num += 1
                wb.save(response)
                return response
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos. %s %s" % (ex,campo1)})

        elif action == 'alumnos_sin_acceso':
            campo1=None
            try:
                __author__ = 'Unemi'
                matriculas = Matricula.objects.filter(status=True, nivel__periodo=periodo, inscripcion__envioemail=False, inscripcion__modalidad__id=3, nivel__sesion_id=13).distinct().order_by('inscripcion__persona__apellido1', 'inscripcion__persona__apellido2','inscripcion__persona__nombres')
                style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
                style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                title = easyxf(
                    'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                style1 = easyxf(num_format_str='D-MMM-YY')
                font_style = XFStyle()
                font_style.font.bold = True
                font_style2 = XFStyle()
                font_style2.font.bold = False
                wb = Workbook(encoding='utf-8')
                ws = wb.add_sheet('alumnos_online_promedio')
                ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                response = HttpResponse(content_type="application/ms-excel")
                response[
                    'Content-Disposition'] = 'attachment; filename=listado_alumnos_asignatura_columna ' + random.randint(
                    1, 10000).__str__() + '.xls'
                columns = [
                    (u"Cedula", 5000),
                    (u"Carrera", 5000),
                    (u"Apellidos", 5000),
                    (u"Nombres", 5000),
                    (u"Usuario", 5000),
                    (u"Institucional", 5000),
                    (u"Correo Personal", 5000),
                    (u"Telefono", 5000),

                    (u"ID asignatura 1", 5000),
                    (u"Asignatura 1", 5000),
                    (u"Codigo Asignatura 1", 5000),
                    (u"Codigo 1", 5000),

                    (u"ID asignatura 2", 5000),
                    (u"Asignatura 2", 5000),
                    (u"Codigo Asignatura 2 ", 5000),
                    (u"Codigo 2 ", 5000),

                    (u"ID asignatura 3", 5000),
                    (u"Asignatura 3", 5000),
                    (u"Codigo Asignatura 3", 5000),
                    (u"Codigo 3", 5000),

                    (u"ID asignatura  4", 5000),
                    (u"Asignatura 4", 5000),
                    (u"Codigo Asignatura 4", 5000),
                    (u"Codigo 4", 5000),

                    (u"ID asignatura  5", 5000),
                    (u"Asignatura 5", 5000),
                    (u"Codigo Asignatura 5", 5000),
                    (u"Codigo 5", 5000),
                ]
                row_num = 4
                for col_num in range(len(columns)):
                    ws.write(row_num, col_num, columns[col_num][0], font_style)
                    ws.col(col_num).width = columns[col_num][1]
                date_format = xlwt.XFStyle()
                date_format.num_format_str = 'yyyy/mm/dd'
                row_num = 5
                for matricula in matriculas:
                    campo1 = str(matricula.inscripcion.persona.identificacion())
                    campo2 = str(matricula.inscripcion.carrera.nombre_completo())
                    campo3 = str(
                        matricula.inscripcion.persona.apellido1 + " " + matricula.inscripcion.persona.apellido2)
                    campo4 = str(matricula.inscripcion.persona.nombres)
                    campo5 = str(
                        matricula.inscripcion.persona.usuario.username if not matricula.inscripcion.persona.usuario == None else "S/N")
                    campo6 = str(matricula.inscripcion.persona.emailinst)
                    campo7 = str(matricula.inscripcion.persona.email)
                    campo8 = str(matricula.inscripcion.persona.telefono)

                    ws.write(row_num, 0, campo1, font_style2)
                    ws.write(row_num, 1, campo2, font_style2)
                    ws.write(row_num, 2, campo3, font_style2)
                    ws.write(row_num, 3, campo4, font_style2)
                    ws.write(row_num, 4, campo5, font_style2)
                    ws.write(row_num, 5, campo6, font_style2)
                    ws.write(row_num, 6, campo7, font_style2)
                    ws.write(row_num, 7, campo8, font_style2)
                    materiasasignadas = matricula.mis_materias_sin_retiro()
                    col_num = 7
                    for materiaasignada in materiasasignadas:
                        materia = materiaasignada.materia
                        paralelo = materiaasignada.materia.paralelo
                        codigo = materia.asignatura.nombre[:2].upper() + "_" + paralelo + '_1S_2019'
                        codigo_paralelo = paralelo + '_1S_2019_' + str(materiaasignada.materia.id)
                        col_num += 1
                        ws.write(row_num, col_num, materiaasignada.materia.id, font_style2)
                        col_num += 1
                        ws.write(row_num, col_num, materia.asignatura.nombre, font_style2)
                        col_num += 1
                        ws.write(row_num, col_num, codigo, font_style2)
                        col_num += 1
                        ws.write(row_num, col_num, codigo_paralelo, font_style2)
                    row_num += 1
                wb.save(response)
                return response
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos. %s %s" % (ex,campo1)})

        elif action == 'estudiantes_sin_acceso_plataforma':
            campo1=None
            try:
                form = ImportarListadoAlumnoForm(request.POST, request.FILES)
                if form.is_valid():
                    nfile = request.FILES['archivo']
                    nfile._name = generar_nombre("importacion_", nfile._name)
                    archivo = Archivo(nombre='IMPORTACION LISTADO ALUMNMOS',
                                      fecha=datetime.now().date(),
                                      archivo=nfile,
                                      tipo_id=ARCHIVO_TIPO_GENERAL)
                    archivo.save()
                    workbook = xlrd.open_workbook(archivo.archivo.file.name)
                    sheet = workbook.sheet_by_index(0)
                    linea = 1
                    modalidad = Modalidad.objects.get(pk=3)
                    for rowx in range(sheet.nrows):
                        if linea > 1:
                            cols = sheet.row_values(rowx)
                            username = str(cols[0]).strip() if cols[0] else None
                            if Persona.objects.filter(usuario__username=username).exists():
                                persona = Persona.objects.filter(usuario__username=username)[0]
                                if Matricula.objects.filter(inscripcion__persona=persona,nivel__periodo=periodo,inscripcion__carrera__modalidad=3).exists():
                                    matricula = Matricula.objects.filter(inscripcion__persona=persona,nivel__periodo=periodo,inscripcion__carrera__modalidad=3)[0]
                                    matricula.inscripcion.envioemail=True
                                    matricula.inscripcion.save()
                        linea += 1
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos. %s %s" % (ex,campo1)})

        elif action == 'addactividadsakai':
            try:
                f = ActividadSakaiForm(request.POST)
                if f.is_valid():
                    inscripcion = Inscripcion.objects.get(pk=int(request.POST['id']))
                    materia = Materia.objects.get(pk=int(request.POST['idmat']))
                    if not ActividadesSakaiAlumno.objects.filter(tipo=f.cleaned_data['tipo'], status=True,
                                                                 inscripcion=inscripcion,
                                                                 materia=materia).exists():
                        actividad = ActividadesSakaiAlumno(inscripcion=inscripcion,materia=materia,
                                                           nombreactividadsakai= f.cleaned_data['nombreactividadsakai'],
                                                           tipo=f.cleaned_data['tipo'],
                                                           nota=f.cleaned_data['nota'],
                                                           notaposible=f.cleaned_data['notaposible']
                        )
                        actividad.save(request)
                    else:
                        raise NameError('Error')
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'asignarcupo':
            try:
                valor = 0
                if Matricula.objects.filter(id=request.POST['matid'], status=True).exists():
                    matricula = Matricula.objects.filter(id=request.POST['matid'], status=True)[0]
                    if matricula.aprobado:
                        matricula.aprobado=False
                        valor =0
                    else:
                        matricula.aprobado = True
                        valor =1
                matricula.save(request)
                log(u'Quito o agrego a un cupo: %s' % matricula, request, "edit")
                return JsonResponse({"result": "ok", "valor": valor})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": "Error al eliminar los datos."})

        elif action == 'reportedistributivo':
            try:
                __author__ = 'Unemi'
                title = easyxf(
                    'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                style1 = easyxf(num_format_str='D-MMM-YY')
                font_style = XFStyle()
                font_style.font.bold = True
                font_style2 = XFStyle()
                font_style2.font.bold = False
                wb = Workbook(encoding='utf-8')
                ws = wb.add_sheet('exp_xls_post_part')
                ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                response = HttpResponse(content_type="application/ms-excel")
                response[
                    'Content-Disposition'] = 'attachment; filename=DISTRIBUTIVO ASIGNATURA ' + periodo.nombre.__str__() + '.xls'
                columns = [
                    (u"FACULTAD", 6000),
                    (u"CARRERA", 6000),
                    (u"MALLA", 6000),
                    (u"SECCIÓN", 6000),
                    (u"NIVEL", 6000),
                    (u"PARALELO", 6000),
                    (u"ASIGNATURA", 6000),
                    (u"TEORICA PRACTICA", 6000),
                    (u"CUPO", 6000),
                    (u"MATRICULADOS", 6000),
                    (u"DOCENTE", 6000),
                    (u"CEDULA", 6000),
                    # (u"USUARIO", 6000),
                    (u"AFINIDAD", 6000),
                    (u"HORAS SEMANALES", 4000),
                    (u"MALLA (HORAS PRESENCIALES SEMANALES)", 4000),
                    (u"TIPO", 4000),
                    (u"CORREO PERSONAL", 4000),
                    (u"CORREO INSTITUCIONAL", 4000),
                    (u"TIPO PROFESOR", 4000),
                    (u"PROFESOR DESDE", 4000),
                    (u"PROFESOR HASTA", 4000),
                    (u"DEDICACION", 5000),
                    (u"CATEGORIA", 4000),
                    (u"INICIO MATERIA", 4000),
                    (u"FIN MATERIA", 4000),
                    (u"FIN ASISTENCIA", 4000),
                    # (u"ID", 4000),
                    (u"TELEFONO", 6000),
                    (u"MODELO EVALUATIVO", 6000),
                    (u"IDMATERIA", 2500),
                    (u"ACEPTACION", 2500),
                    (u"OBSERVACION ACEPTACION", 10000),
                    (u"HORARIO FECHA ACEPTACION", 10000),
                    (u"HORARIO ACEPTACION", 10000),
                    (u"HORARIO OBSERVACION ACEPTACION", 10000),
                    (u"fullname", 10000),
                    (u"shortname", 10000)
                ]
                # periodo = request.GET['periodo']
                row_num = 3
                date_format = xlwt.XFStyle()
                date_format.num_format_str = 'yyyy/mm/dd'
                for col_num in range(len(columns)):
                    ws.write(row_num, col_num, columns[col_num][0], font_style)
                    ws.col(col_num).width = columns[col_num][1]
                cursor = connection.cursor()
                sql = "SELECT " \
                      " sga_coordinacion.nombre AS Facultad, " \
                      " sga_carrera.nombre AS Carrera, " \
                      " sga_sesion.nombre AS Seccion, " \
                      " sga_nivelmalla.nombre AS Nivel, " \
                      " sga_materia.paralelo as Paralelo, " \
                      " sga_materia.id AS Idmateria, " \
                      " sga_asignatura.nombre AS Asignatura," \
                      " sga_persona.apellido1apellido1 || ' ' || sga_persona.apellido2 || ' ' || sga_persona.nombres  AS Docente," \
                      " sga_profesormateria.hora AS sga_profesormateria_hora, " \
                      " (case sga_profesormateria.principal when true then 'PRINCIPAL' else 'PRACTICA' end) as Tipo," \
                      " sga_persona.cedula, (select u.username from auth_user u where u.id=sga_persona.usuario_id), sga_persona.email, sga_persona.emailinst," \
                      " sga_materia.cupo as cupo, (select count(*) from sga_materiaasignada ma, sga_matricula mat1 where ma.matricula_id=mat1.id and mat1.estado_matricula in (2,3) and ma.materia_id=sga_materia.id and ma.id not in (select mr.materiaasignada_id from sga_materiaasignadaretiro mr)) as nmatriculados," \
                      " sga_tipoprofesor.nombre as Tipoprofesor, " \
                      " sga_profesormateria.desde as desde, " \
                      " sga_profesormateria.hasta as hasta, " \
                      " (select ti.nombre from sga_profesordistributivohoras dis,sga_tiempodedicaciondocente ti where dis.dedicacion_id=ti.id and dis.profesor_id=sga_profesor.id and periodo_id=" + str(periodo.id) + " and dis.status=True) as dedicacion," \
                                                                                                                                                                                                               " (select ca.nombre from sga_profesordistributivohoras dis,sga_categorizaciondocente ca where dis.categoria_id=ca.id and dis.profesor_id=sga_profesor.id and dis.periodo_id=" + str(periodo.id) + " and dis.status=True) as categoria, " \
                                                                                                                                                                                                                                                                                                                                                                                                         " (case sga_asignaturamalla.practicas when true then 'SI' else 'NO' end) as tipomateria, " \
                                                                                                                                                                                                                                                                                                                                                                                                         " (case sga_profesormateria.afinidad when true then 'SI' else 'NO' end) as afinidad, " \
                                                                                                                                                                                                                                                                                                                                                                                                         " sga_materia.inicio as inicio, " \
                                                                                                                                                                                                                                                                                                                                                                                                         " sga_materia.fin as fin, " \
                                                                                                                                                                                                                                                                                                                                                                                                         " sga_materia.fechafinasistencias as finasistencia," \
                                                                                                                                                                                                                                                                                                                                                                                                         " sga_materia.id as id, sga_persona.telefono_conv as telefonoconv, sga_persona.telefono as telefono, extract(year from sga_malla.inicio) as anio," \
                                                                                                                                                                                                                                                                                                                                                                                                         " (select modelo.nombre from sga_modeloevaluativo modelo where modelo.id = sga_materia.modeloevaluativo_id) as modeloevaluativo," \
                                                                                                                                                                                                                                                                                                                                                                                                         " sga_asignaturamalla.horaspresencialessemanales as horaspresencialessemanales," \
                                                                                                                                                                                                                                                                                                                                                                                                         " sga_profesormateria.aceptarmateria as aceptarmateria," \
                                                                                                                                                                                                                                                                                                                                                                                                         " sga_profesormateria.aceptarmateriaobs as aceptarmateriaobs, " \
                                                                                                                                                                                                                                                                                                                                                                                                         " sga_profesormateria.fecha_horario as fecha_horario, " \
                                                                                                                                                                                                                                                                                                                                                                                                         " sga_profesormateria.aceptarhorario as aceptarhorario, " \
                                                                                                                                                                                                                                                                                                                                                                                                         " sga_profesormateria.aceptarhorarioobs as aceptarhorarioobs" \
                                                                                                                                                                                                                                                                                                                                                                                                         " FROM public.sga_materia sga_materia" \
                                                                                                                                                                                                                                                                                                                                                                                                         " LEFT JOIN public.sga_profesormateria sga_profesormateria ON sga_materia.id = sga_profesormateria.materia_id and sga_profesormateria.status=true and sga_profesormateria.activo=true" \
                                                                                                                                                                                                                                                                                                                                                                                                         " LEFT JOIN public.sga_profesor sga_profesor ON sga_profesor.id = sga_profesormateria.profesor_id and sga_profesor.status=true " \
                                                                                                                                                                                                                                                                                                                                                                                                         " LEFT JOIN public.sga_tipoprofesor sga_tipoprofesor ON sga_tipoprofesor.id = sga_profesormateria.tipoprofesor_id and sga_tipoprofesor.status=true" \
                                                                                                                                                                                                                                                                                                                                                                                                         " LEFT JOIN public.sga_persona sga_persona ON sga_profesor.persona_id = sga_persona.id and sga_persona.status=true " \
                                                                                                                                                                                                                                                                                                                                                                                                         " INNER JOIN public.sga_nivel sga_nivel ON sga_materia.nivel_id = sga_nivel.id and sga_nivel.status=true " \
                                                                                                                                                                                                                                                                                                                                                                                                         " INNER JOIN public.sga_asignatura sga_asignatura ON sga_materia.asignatura_id = sga_asignatura.id and sga_asignatura.status=true" \
                                                                                                                                                                                                                                                                                                                                                                                                         " INNER JOIN public.sga_asignaturamalla sga_asignaturamalla ON sga_materia.asignaturamalla_id = sga_asignaturamalla.id and sga_asignaturamalla.status=true" \
                                                                                                                                                                                                                                                                                                                                                                                                         " INNER JOIN public.sga_nivelmalla sga_nivelmalla ON sga_asignaturamalla.nivelmalla_id = sga_nivelmalla.id and sga_nivelmalla.status=true " \
                                                                                                                                                                                                                                                                                                                                                                                                         " INNER JOIN public.sga_malla sga_malla ON sga_asignaturamalla.malla_id = sga_malla.id and sga_malla.status=true " \
                                                                                                                                                                                                                                                                                                                                                                                                         " INNER JOIN public.sga_carrera sga_carrera ON sga_malla.carrera_id = sga_carrera.id and sga_carrera.status=true  " \
                                                                                                                                                                                                                                                                                                                                                                                                         " INNER JOIN public.sga_coordinacion_carrera sga_coordinacion_carrera ON sga_carrera.id = sga_coordinacion_carrera.carrera_id" \
                                                                                                                                                                                                                                                                                                                                                                                                         " INNER JOIN public.sga_coordinacion sga_coordinacion ON sga_coordinacion_carrera.coordinacion_id = sga_coordinacion.id and sga_coordinacion.status=true" \
                                                                                                                                                                                                                                                                                                                                                                                                         " INNER JOIN public.sga_sesion sga_sesion ON sga_nivel.sesion_id = sga_sesion.id " \
                                                                                                                                                                                                                                                                                                                                                                                                         " INNER JOIN public.sga_periodo sga_periodo ON sga_nivel.periodo_id = sga_periodo.id and sga_periodo.status=true" \
                                                                                                                                                                                                                                                                                                                                                                                                         " WHERE sga_periodo.id = " + str(periodo.id) + " and sga_materia.status=true" \
                                                                                                                                                                                                                                                                                                                                                                                                                                                " ORDER BY sga_coordinacion.nombre, sga_carrera.nombre, sga_sesion.nombre, sga_nivelmalla.nombre,sga_materia.paralelo,sga_asignatura.nombre"
                cursor.execute(sql)
                results = cursor.fetchall()
                row_num = 4
                for r in results:
                    i = 0
                    campo1 = r[0].__str__()
                    campo2 = r[1].__str__()
                    campo3 = r[2].__str__()
                    campo4 = r[3].__str__()
                    campo5 = r[4].__str__()
                    campo6 = r[5]
                    campo7 = r[6].__str__()
                    campo8 = r[7].__str__()
                    campo9 = r[8]
                    campo10 = r[9].__str__()
                    campo11 = r[10].__str__()
                    # campo12 = r[11]
                    campo13 = r[12].__str__()
                    campo14 = r[13].__str__()
                    campo15 = r[14].__str__()
                    campo16 = r[15].__str__()
                    campo17 = r[16].__str__()
                    campo18 = r[17].__str__()
                    campo19 = r[18].__str__()
                    campo20 = r[19].__str__()
                    campo21 = r[20].__str__()
                    campo22 = r[21].__str__()
                    campo23 = r[22].__str__()
                    campo24 = r[23].__str__()
                    campo25 = r[24].__str__()
                    campo26 = r[25].__str__()
                    # campo27 = r[26]
                    campo28 = r[27].__str__() + " - " + r[28].__str__()
                    campo29 = r[29].__str__()
                    campo30 = r[30].__str__()
                    campo31 = r[31].__str__()
                    if r[33] == None or r[33] == '':
                        campo32 = ''
                        campo33 = ''
                    else:
                        campo32 = 'NO'
                        if r[32]:
                            campo32 = 'SI'
                        campo33 = r[33]
                    campo34 = r[34]
                    if r[36] == None or r[36] == '':
                        campo35 = ''
                        campo36 = ''
                    else:
                        campo35 = 'NO'
                        if r[35]:
                            campo35 = 'SI'
                        campo36 = r[36]

                    campo37= campo7+" ["+campo5+"] "+"-1S 2019"
                    campo38= campo7[:2].upper()+"_"+campo5+'_1S_2019'

                    ws.write(row_num, 0, campo1, font_style2)
                    ws.write(row_num, 1, campo2, font_style2)
                    ws.write(row_num, 2, campo29, font_style2)
                    ws.write(row_num, 3, campo3, font_style2)
                    ws.write(row_num, 4, campo4, font_style2)
                    ws.write(row_num, 5, campo5, font_style2)
                    ws.write(row_num, 6, campo7, font_style2)
                    ws.write(row_num, 7, campo22, font_style2)
                    ws.write(row_num, 8, campo15, font_style2)
                    ws.write(row_num, 9, campo16, font_style2)
                    ws.write(row_num, 10, campo8, font_style2)
                    ws.write(row_num, 11, campo11, font_style2)
                    # ws.write(row_num, 12, campo12, font_style2)
                    ws.write(row_num, 12, campo23, font_style2)
                    ws.write(row_num, 13, campo9, font_style2)
                    ws.write(row_num, 14, campo9, font_style2)
                    ws.write(row_num, 15, campo10, font_style2)
                    ws.write(row_num, 16, campo13, font_style2)
                    ws.write(row_num, 17, campo14, font_style2)
                    ws.write(row_num, 18, campo17, font_style2)
                    ws.write(row_num, 19, campo18, style1)
                    ws.write(row_num, 20, campo19, style1)
                    ws.write(row_num, 21, campo20, style1)
                    ws.write(row_num, 22, campo21, style1)
                    ws.write(row_num, 23, campo24, style1)
                    ws.write(row_num, 24, campo25, style1)
                    ws.write(row_num, 25, campo26, style1)
                    # ws.write(row_num, 25, campo27, font_style2)
                    ws.write(row_num, 26, campo28, font_style2)
                    ws.write(row_num, 27, campo30, font_style2)
                    ws.write(row_num, 28, campo6, font_style2)
                    ws.write(row_num, 29, campo32, font_style2)
                    ws.write(row_num, 30, campo33, font_style2)
                    ws.write(row_num, 31, campo34, date_format)
                    ws.write(row_num, 32, campo35, font_style2)
                    ws.write(row_num, 33, campo36, font_style2)
                    ws.write(row_num, 34, campo37, font_style2)
                    ws.write(row_num, 35, campo38, font_style2)
                    row_num += 1
                wb.save(response)
                return response
            except Exception as ex:
                pass

        elif action== 'importar_listado_alumnos':
            try:
                def fechatope(fecha):
                    contador = 0
                    nuevafecha = fecha
                    # while contador < DIAS_MATRICULA_EXPIRA:
                    while contador < 0:
                        nuevafecha = nuevafecha + timedelta(1)
                        if nuevafecha.weekday() != 5 and nuevafecha.weekday() != 6:
                            contador += 1
                    return nuevafecha

                form = ImportarListadoAlumnoForm(request.POST, request.FILES)
                if form.is_valid():
                    nfile = request.FILES['archivo']
                    nfile._name = generar_nombre("importacion_", nfile._name)
                    archivo = Archivo(nombre='IMPORTACION LISTADO ALUMNMOS',
                                      fecha=datetime.now().date(),
                                      archivo=nfile,
                                      tipo_id=ARCHIVO_TIPO_GENERAL)
                    archivo.save()
                    workbook = xlrd.open_workbook(archivo.archivo.file.name)
                    sheet = workbook.sheet_by_index(0)
                    linea = 1
                    modalidad = Modalidad.objects.get(pk=3)
                    for rowx in range(sheet.nrows):
                        if linea > 1:
                            cols = sheet.row_values(rowx)
                            identificacion = str(cols[0]).strip().upper() if cols[0] else None
                            if not Persona.objects.filter(cedula=identificacion).exists():
                                pais = Pais.objects.get(id=int(cols[8]))
                                provincia = Provincia.objects.get(pk=int(cols[10])) if not cols[10] == '' else None
                                canton = Canton.objects.get(pk=int(cols[12])) if not cols[12] == '' else None

                                persona = Persona(cedula=identificacion,
                                                  apellido1=cols[2],
                                                  apellido2=cols[3],
                                                  nombres=cols[1],
                                                  nacimiento=xlrd.xldate.xldate_as_datetime(cols[16],workbook.datemode).date() if not cols[16] == '' else None,
                                                  pais=pais,
                                                  provincia=provincia,
                                                  canton=canton,
                                                  email=cols[4],
                                                  telefono_conv=cols[15],
                                                  telefono=cols[14])
                                persona.save()
                                username = calculate_username(persona)
                                usuario = generar_usuario(persona, username, ALUMNOS_GROUP_ID)
                                if EMAIL_INSTITUCIONAL_AUTOMATICO:
                                    persona.emailinst = username + '@' + EMAIL_DOMAIN
                                    persona.save()
                                grupo = None
                                if UTILIZA_GRUPOS_ALUMNOS:
                                    grupo = Grupo.objects.get(pk=int(cols[15]))
                                    carrera = grupo.carrera
                                    sesion = grupo.sesion
                                    modalidad = grupo.modalidad
                                    sede = grupo.sede
                                else:
                                    # sesion = Sesion.objects.get(pk=int(cols[12]))
                                    sesion = Sesion.objects.get(pk=13)
                                    carrera = Carrera.objects.get(pk=int(cols[5]))
                                    modalidad = Modalidad.objects.get(pk=3)
                                    # sede = Sede.objects.get(pk=int(cols[15]))
                                    sede = Sede.objects.get(pk=1)
                            else:
                                persona = Persona.objects.filter(cedula=identificacion)[0]

                            if Persona.objects.filter(cedula=identificacion).exists():
                                persona = Persona.objects.filter(cedula=cols[0])[0]
                                sesion = Sesion.objects.get(pk=13)
                                carrera = Carrera.objects.get(pk=int(cols[5]))
                                sede = Sede.objects.get(pk=1)
                                # cordinacion_alias = str(carrera.coordinacion_set.all()[0].alias)
                                cordinacion_alias = 'ADMI'
                                # colegio = persona.inscripcion_set.all()[0].colegio
                                raza_id = 6
                                if not Inscripcion.objects.filter(persona=persona, carrera=carrera).exists():
                                    inscripcion = Inscripcion(persona=persona,
                                                              fecha=datetime.now().date(),
                                                              carrera=carrera,
                                                              modalidad=modalidad,
                                                              sesion=sesion,
                                                              sede=sede,
                                                              )
                                    inscripcion.save()
                                    persona.crear_perfil(inscripcion=inscripcion)
                                    documentos = DocumentosDeInscripcion(inscripcion=inscripcion,
                                                                         titulo=False,
                                                                         acta=False,
                                                                         cedula=False,
                                                                         votacion=False,
                                                                         actaconv=False,
                                                                         partida_nac=False,
                                                                         pre=False,
                                                                         observaciones_pre='',
                                                                         fotos=False)
                                    documentos.save()
                                    preguntasinscripcion = inscripcion.preguntas_inscripcion()
                                    perfil_inscripcion = inscripcion.persona.mi_perfil()

                                    perfil_inscripcion.raza_id = raza_id
                                    perfil_inscripcion.save()
                                    inscripciontesdrive = InscripcionTesDrive(inscripcion=inscripcion,
                                                                              licencia=False,
                                                                              record=False,
                                                                              certificado_tipo_sangre=False,
                                                                              prueba_psicosensometrica=False,
                                                                              certificado_estudios=False)
                                    inscripciontesdrive.save()
                                    # inscripcion.mi_malla()
                                    inscripcion.malla_inscripcion()
                                    inscripcion.actualizar_nivel()
                                    if USA_TIPOS_INSCRIPCIONES:
                                        inscripciontipoinscripcion = InscripcionTipoInscripcion(inscripcion=inscripcion,tipoinscripcion_id=TIPO_INSCRIPCION_INICIAL)
                                        inscripciontipoinscripcion.save()
                                    # persona.creacion_persona(request.session['nombresistema'])
                                else:
                                    inscripcion = Inscripcion.objects.filter(persona=persona, carrera=carrera)[0]
                                    perfil_inscripcion = inscripcion.persona.mi_perfil()
                                    perfil_inscripcion.raza_id = raza_id
                                    perfil_inscripcion.save()

                                inscripcion = Inscripcion.objects.filter(persona=persona, carrera=carrera)[0]
                                inscripcion.sesion = sesion
                                inscripcion.save()

                                # considerar si que existe en titulo persona
                            if not persona.tiene_otro_titulo(inscripcion=inscripcion):
                                nivel = Nivel.objects.get(periodo=periodo, sesion=sesion,paralelo__icontains=cordinacion_alias)
                                # matricula
                                if not inscripcion.matricula_periodo(periodo):
                                    matricula = Matricula(inscripcion=inscripcion,
                                                          nivel=nivel,
                                                          pago=False,
                                                          iece=False,
                                                          becado=False,
                                                          porcientobeca=0,
                                                          fecha=datetime.now().date(),
                                                          hora=datetime.now().time(),
                                                          fechatope=fechatope(datetime.now().date()))
                                    matricula.save()
                                else:
                                    matricula = Matricula.objects.get(inscripcion=inscripcion, nivel=nivel)

                                asignatura_malla = RecordAcademico.objects.values_list('asignatura__id',  flat=True).filter(asignaturamalla__nivelmalla__id__lt=1, aprobada=True)

                                for materia in Materia.objects.filter(nivel__periodo=periodo,paralelo=str(cols[6].strip()),asignaturamalla__malla__carrera=carrera,nivel__sesion=sesion,asignaturamalla__nivelmalla__id=1):
                                    if not MateriaAsignada.objects.filter(matricula=matricula, materia=materia).exists():
                                        matriculas = matricula.inscripcion.historicorecordacademico_set.filter( asignatura=materia.asignatura, fecha__lt=materia.nivel.fin).count() + 1
                                        materiaasignada = MateriaAsignada(matricula=matricula,
                                                                          materia=materia,
                                                                          notafinal=0,
                                                                          asistenciafinal=0,
                                                                          cerrado=False,
                                                                          matriculas=matriculas,
                                                                          observaciones='',
                                                                          # estado_id=NOTA_ESTADO_EN_CURSO)
                                                                          estado_id=3)
                                        materiaasignada.save()
                                        materiaasignada.asistencias()
                                        materiaasignada.evaluacion()
                                        materiaasignada.mis_planificaciones()
                                        materiaasignada.save()
                                matricula.actualizar_horas_creditos()
                                matricula.actualiza_matricula()
                        linea += 1
                log(u'Importacion de listado de alumnos : %s' % personasesion, request, "add")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action== 'importar_listado_soporte_alumno':
            try:
                form = ImportarListadoAlumnoSoporteForm(request.POST, request.FILES)
                if form.is_valid():
                    nfile = request.FILES['archivo']
                    nfile._name = generar_nombre("importacion_", nfile._name)
                    archivo = Archivo(nombre='IMPORTACION LISTADO ALUMNMOS SOPORTES',
                                      fecha=datetime.now().date(),
                                      archivo=nfile,
                                      tipo_id=ARCHIVO_TIPO_GENERAL)
                    archivo.save()
                    workbook = xlrd.open_workbook(archivo.archivo.file.name)
                    sheet = workbook.sheet_by_index(0)
                    linea = 1
                    for rowx in range(sheet.nrows):
                        if linea > 1:
                            cols = sheet.row_values(rowx)
                            identificacion = str(cols[0]).strip().upper() if cols[0] else None
                            idSoporte = int(cols[17]) if cols[17] else None
                            # matricula = Matricula.objects.filter(inscripcion__persona__cedula=identificacion,nivel__periodo=periodo,inscripcion__carrera__modalidad=3)[0]
                            matricula = Matricula.objects.filter(inscripcion__persona__cedula=identificacion,nivel__periodo=periodo)[0]
                            soporte = VirtualSoporteUsuario.objects.get(id=idSoporte)
                            if VirtualSoporteUsuarioInscripcion.objects.filter(status=True, matricula=matricula,soporteusuario=soporte).exists():
                                soporte = VirtualSoporteUsuarioInscripcion.objects.filter(status=True, matricula=matricula,soporteusuario=soporte)[0]
                                if soporte.activo == False:
                                    soporte.activo = True
                                else:
                                    soporte.activo = False
                            else:
                                soporte = VirtualSoporteUsuarioInscripcion(matricula=matricula, soporteusuario=soporte, activo=True)
                            if soporte:
                                soporte.save(request)

                        linea += 1
                log(u'Importacion de listado de alumnos soportes : %s' % personasesion, request, "add")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action== 'importar_listado_soporte_admin':
            try:
                form = ImportarListadoAlumnoSoporteForm(request.POST, request.FILES)
                if form.is_valid():
                    nfile = request.FILES['archivo']
                    nfile._name = generar_nombre("importacion_", nfile._name)
                    archivo = Archivo(nombre='IMPORTACION LISTADO ALUMNMOS SOPORTES',
                                      fecha=datetime.now().date(),
                                      archivo=nfile,
                                      tipo_id=ARCHIVO_TIPO_GENERAL)
                    archivo.save()
                    workbook = xlrd.open_workbook(archivo.archivo.file.name)
                    sheet = workbook.sheet_by_index(0)
                    linea = 1
                    for rowx in range(sheet.nrows):
                        if linea > 1:
                            cols = sheet.row_values(rowx)
                            admin = VirtualSoporteUsuario.objects.get(id=int(cols[1]))
                            soporte = VirtualSoporteUsuario.objects.get(id=int(cols[3]))
                            if  VirtualSoporteAsignado.objects.filter(status=True, soporte=admin,soporteusuario=soporte,periodo=periodo).exists():
                                soporte = VirtualSoporteAsignado.objects.filter(status=True,soporte=admin,soporteusuario=soporte,periodo=periodo)[0]
                                if soporte.activo == False:
                                    soporte.activo = True
                                else:
                                    soporte.activo = False
                            else:
                                soporte = VirtualSoporteAsignado(soporte=admin,soporteusuario=soporte, activo=True,periodo=periodo)
                            if soporte:
                                soporte.save(request)

                        linea += 1
                log(u'Importacion de listado de alumnos soportes : %s' % personasesion, request, "add")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'migrarfechas':
            actividad1=None
            try:
                for actividad in ActividadesSakaiAlumno.objects.filter(status=True).exclude(tipo=4):
                    actividad1=actividad
                    actividad.migar_mi_fecha()
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al actualizar los datos. %s %s" % (ex, actividad1)})

        elif action == 'actualizar_estado_actividades':
            try:
                for matricula in Matricula.objects.filter(status=True,nivel__periodo=periodo,inscripcion__carrera__id=request.POST['idc']).order_by('inscripcion__persona__apellido1','inscripcion__persona__apellido2','inscripcion__persona__nombres'):
                    for actividad in matricula.inscripcion.actividadessakaialumno_set.filter(status=True).exclude(tipo=4):
                        actividad.actualizar_estado_actividades()
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al actualizar los datos. %s %s" })

        elif action == 'informeactividades':
            try:
                __author__ = 'Unemi'
                from itertools import chain
                fini = convertir_fecha_invertida(request.POST['fini'])
                ffin = convertir_fecha_invertida(request.POST['ffin'])
                title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                font_style = XFStyle()
                font_style.font.bold = True
                font_style2 = XFStyle()
                font_style2.font.bold = False
                wb = Workbook(encoding='utf-8')
                ws = wb.add_sheet('actividades')
                ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                response = HttpResponse(content_type="application/ms-excel")
                response['Content-Disposition'] = 'attachment; filename=actividades' + random.randint(1,10000).__str__() + '.xls'
                columns = [
                    (u"cedula", 5000),
                    (u"apellidos", 5000),
                    (u"nombres", 5000),
                    (u"carrera", 5000),
                    (u"asignatura", 5000),
                    (u"tipo", 5000),
                    (u"actividad", 5000),
                    (u"nota obtenida", 5000),
                    (u"nota posible", 5000),
                    (u"fecha inicio", 5000),
                    (u"fecha fin", 5000),
                ]
                row_num = 4
                for col_num in range(len(columns)):
                    ws.write(row_num, col_num, columns[col_num][0], font_style)
                    ws.col(col_num).width = columns[col_num][1]
                date_format = xlwt.XFStyle()
                date_format.num_format_str = 'yyyy/mm/dd'
                row_num = 5
                actividades1 = ActividadesSakaiAlumno.objects.filter(status=True, fechainicio__gte=fini, fechafin__lte=ffin, tipo__in=[1,2],inscripcion__carrera__id=request.POST['idc'])
                actividades2 = ActividadesSakaiAlumno.objects.filter(status=True, fechafin__lte=ffin, tipo=3,inscripcion__carrera__id=request.POST['idc'])
                actividades = list(chain(actividades1, actividades2))
                for actividad in actividades:
                    ws.write(row_num, 0, actividad.inscripcion.persona.identificacion(), font_style2)
                    ws.write(row_num, 1, actividad.inscripcion.persona.apellido1 +" "+actividad.inscripcion.persona.apellido2, font_style2)
                    ws.write(row_num, 2, actividad.inscripcion.persona.nombres, font_style2)
                    ws.write(row_num, 3, actividad.inscripcion.carrera.nombre_completo(), font_style2)
                    ws.write(row_num, 4, actividad.materia.nombre_completo(), font_style2)
                    ws.write(row_num, 5, str(actividad.get_tipo_display()), font_style2)
                    ws.write(row_num, 6, actividad.nombreactividadsakai, font_style2)
                    ws.write(row_num, 7, actividad.nota, font_style2)
                    ws.write(row_num, 8, actividad.notaposible, font_style2)
                    ws.write(row_num, 9, actividad.fechainicio, date_format)
                    ws.write(row_num, 10, actividad.fechafin, date_format)
                    row_num += 1
                wb.save(response)
                return response
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos. %s" % ex})

        elif action == 'informeactividades_pendientes':
            try:
                __author__ = 'Unemi'
                from itertools import chain
                title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                font_style = XFStyle()
                font_style.font.bold = True
                font_style2 = XFStyle()
                font_style2.font.bold = False
                wb = Workbook(encoding='utf-8')
                ws = wb.add_sheet('actividades')
                ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                response = HttpResponse(content_type="application/ms-excel")
                response['Content-Disposition'] = 'attachment; filename=actividades_pendientes' + random.randint(1,10000).__str__() + '.xls'
                columns = [
                    (u"cedula", 5000),
                    (u"apellidos", 5000),
                    (u"nombres", 5000),
                    (u"carrera", 5000),
                    (u"asignatura", 5000),
                    (u"tipo", 5000),
                    (u"actividad", 5000),
                    (u"nota obtenida", 5000),
                    (u"nota posible", 5000),
                    (u"fecha inicio", 5000),
                    (u"fecha fin", 5000),
                ]
                row_num = 4
                for col_num in range(len(columns)):
                    ws.write(row_num, col_num, columns[col_num][0], font_style)
                    ws.col(col_num).width = columns[col_num][1]
                date_format = xlwt.XFStyle()
                date_format.num_format_str = 'yyyy/mm/dd'
                row_num = 5
                # actividades1 = ActividadesSakaiAlumno.objects.filter(status=True, tipo__in=[1, 2], pendiente=True,inscripcion__carrera__id=request.POST['idc']).order_by('inscripcion__persona__apellido1','inscripcion__persona__apellido2')
                # actividades2 = ActividadesSakaiAlumno.objects.filter(status=True, tipo=3, pendiente=True,inscripcion__carrera__id=request.POST['idc']).order_by('inscripcion__persona__apellido1','inscripcion__persona__apellido2')
                # actividades = list(chain(actividades1, actividades2))
                for matricula in Matricula.objects.filter(status=True,nivel__periodo=periodo,inscripcion__carrera__id=request.POST['idc']).distinct().order_by('inscripcion__persona__apellido1','inscripcion__persona__apellido2','inscripcion__persona__nombres'):
                    for actividad in matricula.inscripcion.actividadessakaialumno_set.filter(pendiente=True):
                        ws.write(row_num, 0, actividad.inscripcion.persona.identificacion(), font_style2)
                        ws.write(row_num, 1, actividad.inscripcion.persona.apellido1 +" "+actividad.inscripcion.persona.apellido2, font_style2)
                        ws.write(row_num, 2, actividad.inscripcion.persona.nombres, font_style2)
                        ws.write(row_num, 3, actividad.inscripcion.carrera.nombre_completo(), font_style2)
                        ws.write(row_num, 4, actividad.materia.nombre_completo(), font_style2)
                        ws.write(row_num, 5, str(actividad.get_tipo_display()), font_style2)
                        ws.write(row_num, 6, actividad.nombreactividadsakai, font_style2)
                        ws.write(row_num, 7, actividad.nota, font_style2)
                        ws.write(row_num, 8, actividad.notaposible, font_style2)
                        ws.write(row_num, 9, actividad.fechainicio, date_format)
                        ws.write(row_num, 10, actividad.fechafin, date_format)
                        row_num += 1
                wb.save(response)
                return response
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos. %s" % ex})

        elif action == 'desactivarsoporte':
            try:
                if request.POST['tipo']=='inscripcion':
                    if VirtualSoporteUsuarioInscripcion.objects.filter(id=request.POST['id'], status=True).exists():
                        soporte=VirtualSoporteUsuarioInscripcion.objects.get(id=request.POST['id'], status=True)
                        soporte.activo=False
                        soporte.save(request)
                elif request.POST['tipo']=='tutor':
                    if VirtualSoporteUsuarioProfesor.objects.filter(id=request.POST['id'], activo=True).exists():
                        soporte = VirtualSoporteUsuarioProfesor.objects.get(id=request.POST['id'], activo=True)
                        soporte.activo = False
                        soporte.save(request)
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": "Error al guardar los datos."})

        elif action == 'activarsoporte':
            try:
                soporte = VirtualSoporteUsuario.objects.get(id=request.POST['soporte'])
                if 'matricula' in request.POST:
                    matricula = Matricula.objects.get(id=request.POST['matricula'])
                    if VirtualSoporteUsuarioInscripcion.objects.filter(status=True, matricula=matricula, soporteusuario=soporte).exists():
                        soporte = VirtualSoporteUsuarioInscripcion.objects.filter(status=True, matricula=matricula, soporteusuario=soporte)[0]
                        if soporte.activo == False:
                            soporte.activo = True
                        else:
                            soporte.activo = False
                    else:
                        soporte= VirtualSoporteUsuarioInscripcion(matricula=matricula, soporteusuario=soporte, activo=True )
                    if soporte:
                        soporte.save(request)
                    return JsonResponse({"result": "ok"})

                elif 'tutor' in request.POST:
                    profesor = Profesor.objects.get(id=request.POST['tutor'])
                    if VirtualSoporteUsuarioProfesor.objects.filter(activo=True, profesor=profesor,soporteusuario=soporte).exists():
                        soporte = VirtualSoporteUsuarioProfesor.objects.filter(activo=True, profesor=profesor, soporteusuario=soporte)[0]
                        if soporte.activo == False:
                            soporte.activo = True
                        else:
                            soporte.activo = False
                    else:
                        soporte = VirtualSoporteUsuarioProfesor(profesor=profesor, soporteusuario=soporte,activo=True)
                    if soporte:
                        soporte.save(request)
                    return JsonResponse({"result": "ok"})

            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": "Error al eliminar los datos."})

        elif action == 'activarsoporteadmin':
            try:
                matriculas = Matricula.objects.filter(status=True, nivel__periodo=periodo,inscripcion__modalidad__id=3).distinct().order_by(
                    'inscripcion__persona__apellido1', 'inscripcion__persona__apellido2','inscripcion__persona__nombres')
                soportes = VirtualSoporteUsuario.objects.get(id=request.POST['id'])
                for matricula in matriculas:
                    if VirtualSoporteUsuarioInscripcion.objects.filter(status=True, matricula=matricula, soporteusuario=soportes).exists():
                        soporte = VirtualSoporteUsuarioInscripcion.objects.filter(status=True, matricula=matricula,soporteusuario=soportes)[0]
                        if soporte.activo == False:
                            soporte.activo = True
                    else:
                        soporte = VirtualSoporteUsuarioInscripcion(matricula=matricula, soporteusuario=soportes, activo=True)
                    if soporte:
                        soporte.save(request)
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": "Error al eliminar los datos."})

        elif action == 'asignarestudiantescarrera':
            try:
                form = AsignarSoporteEstudiante(request.POST)
                if form.is_valid():
                    soportes = VirtualSoporteUsuario.objects.get(id=request.POST['soporte'])
                    carrera=form.cleaned_data['carrera']
                    matriculas = Matricula.objects.filter(status=True, nivel__periodo=periodo, inscripcion__carrera__id__in=carrera).distinct().order_by('inscripcion__persona__apellido1', 'inscripcion__persona__apellido2','inscripcion__persona__nombres')
                    for matricula in matriculas:
                        if VirtualSoporteUsuarioInscripcion.objects.filter(status=True, matricula=matricula, soporteusuario=soportes).exists():
                            soporte = VirtualSoporteUsuarioInscripcion.objects.filter(status=True, matricula=matricula,soporteusuario=soportes)[0]
                            if soporte.activo == False:
                                soporte.activo = True
                        else:
                            soporte = VirtualSoporteUsuarioInscripcion(matricula=matricula, soporteusuario=soportes, activo=True)
                        if soporte:
                            soporte.save(request)
                    for eliminar in VirtualSoporteUsuarioInscripcion.objects.filter(status=True, soporteusuario=soportes, matricula__nivel__periodo=periodo).exclude(matricula__inscripcion__carrera__id__in=carrera):
                        eliminar.status=False
                        eliminar.save(request)
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": "Error al eliminar los datos."})

        elif action == 'actualizar_acceso_plataforma':
            try:
                cursor = connections['db_moodle_virtual_select'].cursor()
                query="SELECT distinct b.userid,a.username FROM mooc_logstore_standard_log AS b INNER JOIN mooc_user AS a ON b.userid=a.id AND NOT b.userid IN (7,8,3,4,5,6,1,2)"
                cursor.execute(query)
                results = cursor.fetchall()
                for result in results:
                    if Inscripcion.objects.filter(persona__usuario__username=result[1],carrera__modalidad=3).exists():
                        inscripcion = Inscripcion.objects.filter(persona__usuario__username=result[1])[0]
                        inscripcion.envioemail=True
                        inscripcion.save()

                return JsonResponse({'result': 'ok', 'mensaje': "Alumnos Actualizados"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad',"mensaje": u"Error al actualizar nota."})

        elif action == 'listado_alumnos_online':
            try:
                __author__ = 'Unemi'
                carrera = None
                modalidad = None
                idcarrera = None
                matriculas = Matricula.objects.filter(status=True, nivel__periodo=periodo, inscripcion__carrera__coordinacion__id=9).distinct().order_by('inscripcion__persona__apellido1', 'inscripcion__persona__apellido2', 'inscripcion__persona__nombres')
                if 'modalidad' in request.POST:
                    modalidad = int(request.POST['modalidad'])
                if 'carr' in request.POST:
                    idcarrera = int(request.POST['carr'])
                if idcarrera>0:
                    carrera = Carrera.objects.get(id=idcarrera)
                    matriculas = matriculas.filter(inscripcion__carrera=carrera)
                if modalidad>0:
                    matriculas = matriculas.filter(inscripcion__carrera__modalidad=modalidad)
                title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                font_style = XFStyle()
                font_style.font.bold = True
                font_style2 = XFStyle()
                font_style2.font.bold = False
                wb = Workbook(encoding='utf-8')
                ws = wb.add_sheet('alumnos_admision_online')
                ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                response = HttpResponse(content_type="application/ms-excel")
                response['Content-Disposition'] = 'attachment; filename=alumnos_admision'+random.randint(1,10000).__str__() + '.xls'
                columns = [
                    (u"Cedula .", 5000),
                    (u"Carrera .", 5000),
                    (u"Nombres .", 5000),
                    (u"Apellidos .", 5000),
                    (u"ppl .", 5000),
                    (u"email", 5000),
                    (u"email institucional", 5000),
                    (u"Celular", 5000),
                    (u"Pais residencia", 5000),
                    (u"Provincia residencia", 5000),
                    (u"Cantón residencia", 5000),
                    (u"Parroquia residencia", 5000),
                    (u"Confirmó ficha", 5000),
                    (u"Cantidad evaluaciones", 5000),
                    (u"Discapacidad", 5000),
                    (u"Porcentaje discapacidad", 5000),
                    (u"Carnet", 5000),
                    (u"Deserto", 5000),
                    # (u"Promedio General ", 5000),
                    (u"Estado Estudiante ", 5000),
                    (u"Asignatura", 5000),
                    (u"Paralelo", 5000),
                    (u"Nota Gestión", 5000),
                    (u"Nota examen", 5000),
                    (u"Nota Total", 5000),
                    (u"Estado Asignatura", 5000),
                    (u"Asistencia", 5000),
                    (u"Cant asignaturas", 5000),
                ]
                row_num = 3
                for col_num in range(len(columns)):
                    ws.write(row_num, col_num, columns[col_num][0], font_style)
                    ws.col(col_num).width = columns[col_num][1]
                date_format = xlwt.XFStyle()
                date_format.num_format_str = 'yyyy/mm/dd'
                row_num = 4
                row_numa = 0
                for matricula in matriculas:
                    campo1 = str(matricula.inscripcion.persona.identificacion())
                    campo2 = str(matricula.inscripcion.carrera.nombre_completo())
                    campo3 = str(matricula.inscripcion.persona.nombres)
                    campo4 = str(matricula.inscripcion.persona.apellido1 + " " + matricula.inscripcion.persona.apellido2)
                    campo5 = "SI" if matricula.inscripcion.persona.ppl else "NO"
                    campo6 = str(matricula.inscripcion.persona.email)
                    campo7 = str(matricula.inscripcion.persona.emailinst)
                    campo8 = str(matricula.inscripcion.persona.telefono) if matricula.inscripcion.persona.telefono else "S/N"
                    campo9 = str(matricula.inscripcion.persona.pais) if matricula.inscripcion.persona.pais else "S/N"
                    campo10 = str( matricula.inscripcion.persona.provincia) if matricula.inscripcion.persona.provincia else "S/N"
                    campo11 = str(matricula.inscripcion.persona.canton) if matricula.inscripcion.persona.canton else "S/N"
                    campo12 = str(matricula.inscripcion.persona.parroquia) if matricula.inscripcion.persona.parroquia else "S/N"
                    campo13 = str("SI") if matricula.inscripcion.persona.tiene_ficha_confirmada() else "NO"
                    campo14 = matricula.cantidad_evaluacion_estudiantes_realizada_periodo(periodo)
                    campo15 = str( matricula.inscripcion.persona.mi_perfil().tipodiscapacidad) if matricula.inscripcion.persona.mi_perfil().tipodiscapacidad else "S/N"
                    campo16 = str( matricula.inscripcion.persona.mi_perfil().porcientodiscapacidad) if matricula.inscripcion.persona.mi_perfil().porcientodiscapacidad else "0"
                    campo17 = str( matricula.inscripcion.persona.mi_perfil().carnetdiscapacidad) if matricula.inscripcion.persona.mi_perfil().carnetdiscapacidad else "S/N"
                    campo18 = "SI" if matricula.inscripcion.desertaonline else "NO"
                    ws.write(row_num, 0, campo1, font_style2)
                    ws.write(row_num, 1, campo2, font_style2)
                    ws.write(row_num, 2, campo3, font_style2)
                    ws.write(row_num, 3, campo4, font_style2)
                    ws.write(row_num, 4, campo5, font_style2)
                    ws.write(row_num, 5, campo6, font_style2)
                    ws.write(row_num, 6, campo7, font_style2)
                    ws.write(row_num, 7, campo8, font_style2)
                    ws.write(row_num, 8, campo9, font_style2)
                    ws.write(row_num, 9, campo10, font_style2)
                    ws.write(row_num, 10, campo11, font_style2)
                    ws.write(row_num, 11, campo12, font_style2)
                    ws.write(row_num, 12, campo13, font_style2)
                    ws.write(row_num, 13, campo14, font_style2)
                    ws.write(row_num, 14, campo15, font_style2)
                    ws.write(row_num, 15, campo16, font_style2)
                    ws.write(row_num, 16, campo17, font_style2)
                    ws.write(row_num, 17, campo18, font_style2)
                    estado_asignatura = "APROBADA"
                    nota_gestion = 0
                    nota_examen = 0
                    gestionyexamen_migrada = 0
                    estado_estudiante="APROBADO"
                    materiasasignadas = matricula.materiaasignada_set.filter(status=True,
                                                                             materiaasignadaretiro__isnull=True,
                                                                             materia__nivel__periodo=periodo,
                                                                             materia__esintroductoria=False).distinct()

                    # campo25 = matricula.inscripcion.porcentaje_total_estudiante_sobre_70_migrada(periodo)
                    # ws.write(row_num, 18, campo25, font_style2)
                    # for materiaasignada in materiasasignadas:
                    #     notafinal = matricula.inscripcion.gestionyexamen_migrada(materiaasignada.materia.id)
                    #     if notafinal<70:
                    #         estado_estudiante = "REPROBADO"
                    aux = 0
                    for materiaasignada in materiasasignadas:
                        estado_asignatura = "APROBADA"
                        nota_gestion=0
                        nota_examen=0
                        gestionyexamen_migrada=0
                        campo20 = str(materiaasignada.materia.asignatura)
                        campo21 = str(materiaasignada.materia.paralelo)
                        for notasmooc in materiaasignada.materia.notas_de_moodle(matricula.inscripcion.persona):
                            nota_gestion+=null_to_decimal(notasmooc[0],0)

                        # nota_gestion = matricula.inscripcion.promedio_general_por_asignatura_migradas1(materiaasignada.materia.id)
                        asistenciafinal = str(materiaasignada.asistenciafinal)
                        numeroasignaturas = str(materiasasignadas.count())
                        nota_examen = matricula.inscripcion.nota_examen_migrada(materiaasignada.materia.id) if matricula.inscripcion.nota_examen_migrada(materiaasignada.materia.id) else 0
                        # gestionyexamen_migrada = matricula.inscripcion.gestionyexamen_migrada(materiaasignada.materia.id)
                        gestionyexamen_migrada = null_to_decimal(nota_gestion + nota_examen,0)
                        if gestionyexamen_migrada<70:
                            estado_asignatura="REPROBADA"
                            estado_estudiante = "REPROBADO"

                        ws.write(row_num, 19, campo20, font_style2)
                        ws.write(row_num, 20, campo21 if campo21 else 0, font_style2)
                        ws.write(row_num, 21, nota_gestion if nota_gestion else 0, font_style2)
                        ws.write(row_num, 22, nota_examen if nota_examen else 0, font_style2)
                        ws.write(row_num, 23, gestionyexamen_migrada if gestionyexamen_migrada else 0, font_style2)
                        ws.write(row_num, 24, estado_asignatura if estado_asignatura else "", font_style2)
                        ws.write(row_num, 25, asistenciafinal if asistenciafinal else 0, font_style2)
                        ws.write(row_num, 26, numeroasignaturas if numeroasignaturas else 0, font_style2)
                        if aux==0:
                            row_numa=row_num
                        aux+=1
                        row_num += 1
                    ws.write(row_numa, 18, estado_estudiante, font_style2)
                wb.save(response)
                return response
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos. %s" % ex})

        elif action == 'listado_alumnos_gestion1':
            try:
                __author__ = 'Unemi'
                carrera = None
                modalidad = None
                idcarrera = None
                nota_gestion = 0
                notaexamen = 0
                notafinal = 0
                matriculas = Matricula.objects.filter(status=True, nivel__periodo=periodo, inscripcion__carrera__coordinacion__id=9).distinct().order_by('inscripcion__persona__apellido1', 'inscripcion__persona__apellido2', 'inscripcion__persona__nombres')
                if 'modalidad' in request.POST:
                    modalidad = int(request.POST['modalidad'])
                if 'carr' in request.POST:
                    idcarrera = int(request.POST['carr'])
                if idcarrera>0:
                    carrera = Carrera.objects.get(id=idcarrera)
                    matriculas = matriculas.filter(inscripcion__carrera=carrera)
                if modalidad>0:
                    matriculas = matriculas.filter(inscripcion__carrera__modalidad=modalidad)
                title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                font_style = XFStyle()
                font_style.font.bold = True
                font_style2 = XFStyle()
                font_style2.font.bold = False
                wb = Workbook(encoding='utf-8')
                ws = wb.add_sheet('alumnos_admision_online')
                ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                response = HttpResponse(content_type="application/ms-excel")
                response['Content-Disposition'] = 'attachment; filename=alumnos_admision'+random.randint(1,10000).__str__() + '.xls'
                columns = [
                    (u"Cedula .", 5000),
                    (u"Carrera .", 5000),
                    (u"Nombres .", 5000),
                    (u"Apellidos .", 5000),
                    (u"Asignatura .", 5000),
                    (u"Nota final", 5000),
                    (u"Estado estudiante", 5000),
                    (u"Estado materia", 5000),
                    (u"Nota Gestión", 5000),
                    (u"Nota Examen", 5000),

                ]
                row_num = 3
                for col_num in range(len(columns)):
                    ws.write(row_num, col_num, columns[col_num][0], font_style)
                    ws.col(col_num).width = columns[col_num][1]
                date_format = xlwt.XFStyle()
                date_format.num_format_str = 'yyyy/mm/dd'
                row_num = 4
                for matricula in matriculas:
                    estado_estudiante = "APROBADO"
                    campo1 = str(matricula.inscripcion.persona.identificacion())
                    campo2 = str(matricula.inscripcion.carrera.nombre_completo())
                    campo3 = str(matricula.inscripcion.persona.nombres)
                    campo4 = str(matricula.inscripcion.persona.apellido1 + " " + matricula.inscripcion.persona.apellido2)
                    nota_gestion = 0
                    notaexamen = 0
                    notafinal = 0
                    materiasasignadas = matricula.materiaasignada_set.filter(status=True,
                                                                             materiaasignadaretiro__isnull=True,
                                                                             materia__nivel__periodo=periodo,
                                                                             materia__esintroductoria=False).distinct()

                    for materiaasignada in materiasasignadas:
                        notafinal = materiaasignada.notafinal
                        estado_materia = str(materiaasignada.estado)
                        if EvaluacionGenerica.objects.filter(materiaasignada=materiaasignada).exists():
                            evaluacion = EvaluacionGenerica.objects.filter(materiaasignada=materiaasignada)
                            notaexamen = evaluacion.filter(detallemodeloevaluativo=71)[0].valor if evaluacion.filter(detallemodeloevaluativo=71) else 0
                            nota_gestion = evaluacion.filter(detallemodeloevaluativo__id__in=[69,68,70]).aggregate(total=Sum('valor'))['total'] if evaluacion.filter(detallemodeloevaluativo__id__in=[69,68,70]) else 0
                        if notafinal <70:
                            estado_estudiante="REPROBADO"
                        ws.write(row_num, 0, campo1, font_style2)
                        ws.write(row_num, 1, campo2, font_style2)
                        ws.write(row_num, 2, campo3, font_style2)
                        ws.write(row_num, 3, campo4, font_style2)
                        ws.write(row_num, 4, str(materiaasignada.materia), font_style2)
                        ws.write(row_num, 5, notafinal if notafinal else 0, font_style2)
                        ws.write(row_num, 6, estado_estudiante, font_style2)
                        ws.write(row_num, 7, estado_materia, font_style2)
                        ws.write(row_num, 8, nota_gestion, font_style2)
                        ws.write(row_num, 9, notaexamen, font_style2)
                        row_num += 1
                wb.save(response)
                return response
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos. %s" % ex})

        elif action == 'listado_alumnos_gestion':
            try:
                __author__ = 'Unemi'
                title = easyxf( 'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                font_style = XFStyle()
                font_style.font.bold = True
                font_style2 = XFStyle()
                font_style2.font.bold = False
                wb = Workbook(encoding='utf-8')
                response = HttpResponse(content_type="application/ms-excel")
                response[ 'Content-Disposition'] = 'attachment; filename=reporte_asignaturas_estudiantes_sin_profesor' + random.randint(1, 10000).__str__() + '.xls'
                cursor = connections['sga_select'].cursor()
                columns = [
                    (u"FACULTAD", 6000),
                    (u"CARRERA", 6000),
                    (u"CEDULA", 3000),
                    (u"PRIMER APELLIDO", 6000),
                    (u"SEGUNDO APELLIDO", 6000),
                    (u"NOMBRES", 6000),
                    (u"EMAIL", 6000),
                    (u"EMAIL INSTITUCIONAL", 6000),
                    (u"CELULAR", 3000),
                    (u"TELEFONO", 3000),
                    (u"MATERIA", 6000),
                    (u"PARALELO", 3000),
                    (u"NIVEL", 3000),
                    (u"NOTA FINAL", 6000),
                    (u"ASISTENCIA FINAL", 6000),
                    (u"ESTADO", 3000),
                    (u"CATEGORIA", 3000),
                    (u"VALOR", 3000),
                ]
                sql = """
                                    SELECT DISTINCT coordinacion.nombre AS coordinacion, CASE WHEN carrera.mencion = '' THEN carrera.nombre ELSE (carrera.nombre || ' CON MENCION EN ' || carrera.mencion) END AS carrera, 
                                 persona.cedula, persona.apellido1, persona.apellido2, persona.nombres, persona.email, persona.emailinst, persona.telefono, persona.telefono_conv, 
                                 (asignatura.nombre || CASE WHEN materia.alias = '' THEN '' ELSE (' - ('|| materia.alias || ')') END || CASE WHEN materia.identificacion = '' THEN '' ELSE (' - ['|| materia.identificacion || ']') END) AS materia, 
                                 materia.paralelo AS paralelo, nivelmalla.nombre AS nivel, materiaasig.notafinal AS notafinal,
                                  materiaasig.asistenciafinal AS asistenciafinal, tipoestado.nombre AS estado,  detmod.nombre, eval.valor
                                FROM sga_materiaasignada materiaasig
                                INNER JOIN sga_evaluaciongenerica eval ON eval.materiaasignada_id=materiaasig.id
                                INNER JOIN sga_detallemodeloevaluativo detmod ON detmod.id=eval.detallemodeloevaluativo_id
                                INNER JOIN sga_materia materia ON materia.id = materiaasig.materia_id
                                INNER JOIN sga_nivel nivel ON nivel.id = materia.nivel_id
                                INNER JOIN sga_asignaturamalla asigmalla ON asigmalla.id = materia.asignaturamalla_id
                                INNER JOIN sga_malla malla ON malla.id = asigmalla.malla_id
                                INNER JOIN sga_carrera carrera ON carrera.id = malla.carrera_id
                                INNER JOIN sga_matricula matricula ON matricula.id = materiaasig.matricula_id
                                INNER JOIN sga_inscripcion inscripcion ON inscripcion.id = matricula.inscripcion_id
                                INNER JOIN sga_persona persona ON persona.id = inscripcion.persona_id
                                INNER JOIN sga_asignatura asignatura ON asignatura.id = materia.asignatura_id
                                INNER JOIN sga_nivelmalla nivelmalla ON nivelmalla.id = asigmalla.nivelmalla_id
                                INNER JOIN sga_tipoestado tipoestado ON tipoestado.id = materiaasig.estado_id
                                INNER JOIN sga_coordinacion_carrera coordinacioncarrera ON coordinacioncarrera.carrera_id = carrera.id
                                INNER JOIN sga_coordinacion coordinacion ON coordinacion.id = coordinacioncarrera.coordinacion_id
                                WHERE coordinacion.id IN(9) AND nivel.periodo_id = %s AND detmod.id IN (69,68,70, 71) AND carrera.modalidad=3 AND materia.esintroductoria = false limit 100  
                                """ % periodo.id.__str__()
                cursor.execute(sql)
                results = cursor.fetchall()
                ws=None
                i = 0
                aux=[]
                for r in results:
                    if i== 0:
                        ws = wb.add_sheet('exp_xls_post_part1')
                        ws.write_merge(0, 0, 0, 7, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                        row_num = 3
                        for col_num in range(len(columns)):
                            ws.write(row_num, col_num, columns[col_num][0], font_style)
                            ws.col(col_num).width = columns[col_num][1]
                        row_num = 4
                    i+=1
                    campo11 = r[10]
                    if not campo11 in aux:
                        aux.append(campo11)
                        row_num += 1
                        columna1 = 15
                        columna2 = 16
                        campo1 = r[0]
                        campo2 = r[1]
                        campo3 = r[2]
                        campo4 = r[3]
                        campo5 = r[4]
                        campo6 = r[5]
                        campo7 = r[6]
                        campo8 = r[7]
                        campo9 = r[8]
                        campo10 = r[9]
                        campo12 = r[11]
                        campo13 = r[12]
                        campo14 = r[13]
                        campo15 = r[14]
                        campo16 = r[15]

                        ws.write(row_num, 0, campo1.__str__(), font_style2)
                        ws.write(row_num, 1, campo2.__str__(), font_style2)
                        ws.write(row_num, 2, campo3.__str__(), font_style2)
                        ws.write(row_num, 3, campo4.__str__(), font_style2)
                        ws.write(row_num, 4, campo5.__str__(), font_style2)
                        ws.write(row_num, 5, campo6.__str__(), font_style2)
                        ws.write(row_num, 6, campo7.__str__(), font_style2)
                        ws.write(row_num, 7, campo8.__str__(), font_style2)
                        ws.write(row_num, 8, campo9.__str__(), font_style2)
                        ws.write(row_num, 9, campo10.__str__(), font_style2)
                        ws.write(row_num, 10, campo11.__str__(), font_style2)
                        ws.write(row_num, 11, campo12.__str__(), font_style2)
                        ws.write(row_num, 12, campo13.__str__(), font_style2)
                        ws.write(row_num, 13, campo14.__str__(), font_style2)
                        ws.write(row_num, 14, campo15.__str__(), font_style2)
                        ws.write(row_num, 15, campo16.__str__(), font_style2)
                    campo17 = r[16]
                    campo18 = r[17]
                    columna1 += 1
                    columna2 += 1
                    ws.write(row_num, columna1, campo17.__str__(), font_style2)
                    ws.write(row_num, columna2, campo18.__str__(), font_style2)
                wb.save(response)
                return response
            except Exception as ex:
                pass

        elif action == 'auditoria':
            try:
                baseDate = datetime.today()
                year = request.POST['year'] if 'year' in request.POST and request.POST['year'] else baseDate.year
                month = request.POST['month'] if 'month' in request.POST and request.POST['month'] else baseDate.month
                data['idi'] = request.POST['id']
                data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=int(encrypt(request.POST['id'])))
                logs = LogEntry.objects.filter(Q(change_message__icontains=inscripcion.persona.__str__()) | Q(user=inscripcion.persona.usuario), action_time__year=year).exclude(user__is_superuser=True)
                logs1 = LogEntryBackup.objects.filter(Q(change_message__icontains=inscripcion.persona.__str__()) | Q(user=inscripcion.persona.usuario), action_time__year=year).exclude(user__is_superuser=True)
                logs2 = LogEntryBackupdos.objects.filter(Q(change_message__icontains=inscripcion.persona.__str__()) | Q(user=inscripcion.persona.usuario), action_time__year=year).exclude(user__is_superuser=True)
                logs3 = LogEntryLogin.objects.filter(user=inscripcion.persona.usuario, action_time__year=year).exclude(user__is_superuser=True, action_app=2)
                addmaterias = AgregacionEliminacionMaterias.objects.filter(matricula__inscripcion=inscripcion, fecha__year=year).order_by('-fecha')
                if int(month):
                    logs = logs.filter(action_time__month=month)
                    logs1 = logs1.filter(action_time__month=month)
                    logs2 = logs2.filter(action_time__month=month)
                    logs3 = logs3.filter(action_time__month=month)
                    addmaterias = addmaterias.filter(fecha__month=month)
                logslist0 = list(logs.values_list("action_time", "action_flag", "change_message", "user__username"))
                logslist1 = list(logs1.values_list("action_time", "action_flag", "change_message", "user__username"))
                logslist2 = list(logs2.values_list("action_time", "action_flag", "change_message", "user__username"))
                logslist = logslist0 + logslist1 + logslist2
                aLogList = []
                for xItem in logslist:
                    # print(xItem)
                    if xItem[1] == 1:
                        action_flag = '<label class="label label-success">AGREGAR</label>'
                    elif xItem[1] == 2:
                        action_flag = '<label class="label label-info">EDITAR</label>'
                    elif xItem[1] == 3:
                        action_flag = '<label class="label label-important">ELIMINAR</label>'
                    else:
                        action_flag = '<label class="label label-warning">OTRO</label>'
                    aLogList.append({"action_time": xItem[0],
                                     "action_flag": action_flag,
                                     "change_message": xItem[2],
                                     "username": xItem[3]})
                for xItem in list(logs3.values_list("action_time", "action_flag", "change_message", "user__username")):
                    if xItem[1] == 1:
                        action_flag = '<label class="label label-success">EXITOSO</label>'
                    elif xItem[1] == 2:
                        action_flag = '<label class="label label-warning">FALLIDO</label>'
                    else:
                        action_flag = '<label class="label label-important">DESCONOCIDO</label>'
                    aLogList.append({"action_time": xItem[0],
                                     "action_flag": action_flag,
                                     "change_message": xItem[2] if xItem[2] else '',
                                     "username": xItem[3]
                                     })
                addmateriaslist = list(addmaterias.values_list("fecha", "agregacion", "asignatura__nombre",
                                                               "responsable__usuario__username"))
                aLogAddMateriaslist = []
                my_time = datetime.min.time()
                for xItem in addmateriaslist:
                    # print(xItem)
                    aLogAddMateriaslist.append({"action_time": datetime.combine(xItem[0], my_time),
                                                "action_flag": ADDITION if xItem[1] else DELETION,
                                                "change_message": u"%s la asignatura %s" % (
                                                ("Agrego" if xItem[1] else "Elimino"), xItem[2]),
                                                "username": xItem[3]})
                datalogs = aLogList + aLogAddMateriaslist
                data['logs'] = sorted(datalogs, key=lambda x: x['action_time'], reverse=True)
                numYear = 6
                dateListYear = []
                for x in range(0, numYear):
                    dateListYear.append((baseDate.year) - x)
                data['list_years'] = dateListYear
                data['year_now'] = int(year)
                data['month_now'] = int(month)
                template = get_template('inscripciones_admision/auditoria.html')
                json_contenido = template.render(data)
                return JsonResponse({"result": "ok", "contenido": json_contenido})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al consultar los datos."})

        elif action == 'delete-cache':
            try:
                try:
                    eInscripcion = Inscripcion.objects.get(pk=int(encrypt(request.POST['id'])))
                except ObjectDoesNotExist:
                    raise NameError(u"No se encontro inscripción")
                ePersona = eInscripcion.persona
                ePersona.delete_cache()
                eInscripcion.delete_cache()
                for eMatricula in Matricula.objects.filter(inscripcion=eInscripcion):
                    eMatricula.delete_cache()
                return JsonResponse({"result": "ok", "mensaje": u""})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": ex.__str__()})

        elif action == 'addppl':
            try:
                f = PersonaPPLForm(request.POST, request.FILES)
                archivoppl = None
                if 'archivoppl' in request.FILES:
                    arch = request.FILES['archivoppl']
                    extension = arch._name.split('.')
                    tam = len(extension)
                    exte = extension[tam - 1]
                    if arch.size > 4194304:
                        raise NameError(u"Tamaño del archivo es mayor a 4 Mb")
                    if not exte.lower() == 'pdf':
                        raise NameError(u"Solo se permiten archivos .pdf")
                    archivoppl = request.FILES['archivoppl']
                    archivoppl._name = generar_nombre("archivoppl_", archivoppl._name)
                if not 'idi' in request.POST:
                    raise NameError(u"Inscripción no encontrada")
                if not Inscripcion.objects.filter(pk=int(request.POST['idi'])).exists():
                    raise NameError(u"Inscripción no encontrada")
                inscripcion = Inscripcion.objects.get(pk=int(request.POST['idi']))

                if not f.is_valid():
                    raise NameError('Formulario incorrecto')

                if HistorialPersonaPPL.objects.filter(persona=inscripcion.persona, fechaingreso=f.cleaned_data['fechaingresoppl']).exists():
                    raise NameError(u"Registro ya existe")

                historialppl = HistorialPersonaPPL(persona=inscripcion.persona,
                                                   observacion=f.cleaned_data['observacionppl'] if f.cleaned_data['observacionppl'] else None,
                                                   archivo=archivoppl,
                                                   fechaingreso=f.cleaned_data['fechaingresoppl'],
                                                   fechasalida=f.cleaned_data['fechasalidappl'] if f.cleaned_data['fechasalidappl'] else None,
                                                   centrorehabilitacion=f.cleaned_data['centrorehabilitacion'] if f.cleaned_data['centrorehabilitacion'] else None,
                                                   lidereducativo=f.cleaned_data['lidereducativo'] if f.cleaned_data['lidereducativo'] else None,
                                                   correolidereducativo=f.cleaned_data['correolidereducativo'] if f.cleaned_data['correolidereducativo'] else None,
                                                   telefonolidereducativo=f.cleaned_data['telefonolidereducativo'] if f.cleaned_data['telefonolidereducativo'] else None,
                                                   )
                historialppl.save(request)
                log(u'Adiciono registro PPL al estudiante: %s' % historialppl, request, "add")
                messages.add_message(request, messages.SUCCESS, f'Se guardo correctamente el registro')
                return JsonResponse({"result": "ok"})

            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos. %s" % ex.__str__()})

        elif action == 'editppl':
            try:
                if not HistorialPersonaPPL.objects.filter(pk=int(request.POST['id'])).exists():
                    raise NameError(u"Registro no existe")
                hppl = HistorialPersonaPPL.objects.get(pk=int(request.POST['id']))
                archivoppl = None
                f = PersonaPPLForm(request.POST, request.FILES)
                if 'archivoppl' in request.FILES:
                    arch = request.FILES['archivoppl']
                    extension = arch._name.split('.')
                    tam = len(extension)
                    exte = extension[tam - 1]
                    if arch.size > 4194304:
                        raise NameError(u"Tamaño del archivo es mayor a 4 Mb")
                    if not exte.lower() == 'pdf':
                        raise NameError(u"Solo se permiten archivos .pdf")
                    archivoppl = request.FILES['archivoppl']
                    archivoppl._name = generar_nombre("archivoppl_", archivoppl._name)

                if not f.is_valid():
                    raise NameError('Formulario incorrecto')

                if HistorialPersonaPPL.objects.filter(persona=hppl.persona, fechaingreso=f.cleaned_data['fechaingresoppl']).exclude(pk=hppl.id).exists():
                    raise NameError(u"Registro ya existe")
                hppl.fechaingreso =f.cleaned_data['fechaingresoppl']
                hppl.fechasalida = f.cleaned_data['fechasalidappl'] if f.cleaned_data['fechasalidappl'] else None
                hppl.centrorehabilitacion = f.cleaned_data['centrorehabilitacion'] if f.cleaned_data['centrorehabilitacion'] else None
                hppl.lidereducativo = f.cleaned_data['lidereducativo'] if f.cleaned_data['lidereducativo'] else None
                hppl.correolidereducativo = f.cleaned_data['correolidereducativo'] if f.cleaned_data['correolidereducativo'] else None
                hppl.telefonolidereducativo = f.cleaned_data['telefonolidereducativo'] if f.cleaned_data['telefonolidereducativo'] else None
                if archivoppl:
                    hppl.archivo = archivoppl
                hppl.observacion = f.cleaned_data['observacionppl'] if f.cleaned_data['observacionppl'] else None
                hppl.save(request)
                log(u'Edito registro PPL al estudiante: %s' % hppl, request, "edit")
                messages.add_message(request, messages.SUCCESS, f'Se guardo correctamente el registro')
                return JsonResponse({"result": "ok"})

            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos. %s" % ex.__str__()})

        elif action == 'deleteppl':
            try:
                puede_realizar_accion(request, 'sga.puede_eliminar_ppl')
                hppl = HistorialPersonaPPL.objects.get(pk=request.POST['id'])
                hppl.delete()
                log(u'Elimino registro de ppl: %s' % hppl, request, "del")
                messages.add_message(request, messages.SUCCESS, f'Se elimino correctamente el registro')
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'traeralumnosmoodle':
            try:
                materia = Materia.objects.get(pk=request.POST['id'], status=True)
                estudiantes = materia.asignados_a_esta_materia_moodle().filter(retiramateria=False)
                primerestudiante = estudiantes.filter(matricula__bloqueomatricula=False).first()
                bandera = True
                modelo_mood = ''
                modelo_sga = ''
                for notasmooc in materia.notas_de_moodle(primerestudiante.matricula.inscripcion.persona):
                    bandera = primerestudiante.evaluacion_generica().filter(detallemodeloevaluativo__nombre=notasmooc[1].upper()).exists()
                    if not bandera:
                        for notasmoocstr in materia.notas_de_moodle(primerestudiante.matricula.inscripcion.persona):
                            modelo_mood += "{}, ".format(notasmoocstr[1])
                        for notassga in primerestudiante.evaluacion_generica():
                            modelo_sga += "{}, ".format(notassga.detallemodeloevaluativo.nombre)
                        return JsonResponse({"result": "bad", "mensaje": u"Modelo Evaluativo extraido es diferente al modelo existente\nMoodle:\n{}\nSGA:\n{}".format(modelo_mood, modelo_sga)})
                listaenviar = estudiantes.filter(matricula__bloqueomatricula=False).values('id', 'matricula__inscripcion__persona__apellido1', 'matricula__inscripcion__persona__apellido2', 'matricula__inscripcion__persona__nombres').order_by('matricula__inscripcion__persona__apellido1')
                return JsonResponse({"result": "ok", "cantidad": len(listaenviar), "inscritos": convertir_lista(listaenviar)})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'traernotaindividual':
            try:
                materia = Materia.objects.get(pk=request.POST['idmateria'], status=True)
                alumno = MateriaAsignada.objects.get(pk=request.POST['id'])
                if materia.notas_de_moodle(alumno.matricula.inscripcion.persona):
                    for notasmooc in materia.notas_de_moodle(alumno.matricula.inscripcion.persona):
                        campo = alumno.campo(notasmooc[1].upper())
                        # if not alumno.matricula.bloqueomatricula:
                        if type(notasmooc[0]) is Decimal:
                            if null_to_decimal(campo.valor) != float(notasmooc[0]) or (
                                    alumno.asistenciafinal < campo.detallemodeloevaluativo.modelo.asistenciaaprobar):
                                actualizar_nota_planificacion(alumno.id, notasmooc[1].upper(), notasmooc[0])
                                auditorianotas = AuditoriaNotas(evaluaciongenerica=campo, manual=False,
                                                                calificacion=notasmooc[0])
                                auditorianotas.save(request)
                        else:
                            if null_to_decimal(campo.valor) != float(0) or (
                                    alumno.asistenciafinal < campo.detallemodeloevaluativo.modelo.asistenciaaprobar):
                                actualizar_nota_planificacion(alumno.id, notasmooc[1].upper(), notasmooc[0])
                                auditorianotas = AuditoriaNotas(evaluaciongenerica=campo, manual=False, calificacion=0)
                                auditorianotas.save(request)
                else:
                    for detallemodelo in materia.modeloevaluativo.detallemodeloevaluativo_set.filter(migrarmoodle=True):
                        campo = alumno.campo(detallemodelo.nombre)
                        actualizar_nota_planificacion(alumno.id, detallemodelo.nombre, 0)
                        auditorianotas = AuditoriaNotas(evaluaciongenerica=campo, manual=False, calificacion=0)
                        auditorianotas.save(request)
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos. %s" % ex})

        elif action == 'cerrarmateria':
            try:
                materia = Materia.objects.get(pk=int(encrypt(request.POST['id'])))
                materia.cerrado = True
                materia.fechacierre = datetime.now().date()
                materia.save(request)
                for asig in materia.asignados_a_esta_materia():
                    asig.cerrado = True
                    asig.save(request, actualiza=False)
                    asig.actualiza_estado()
                for asig in materia.asignados_a_esta_materia():
                    asig.cierre_materia_asignada()
                log(u'Cerro la materia: %s' % materia, request, "add")
                # send_html_mail("Cierre de materia", "emails/cierremateria.html", {'profesor': profesor, 'materia': materia, 't': miinstitucion()}, profesor.persona.lista_emails_envio(), [], cuenta=CUENTAS_CORREOS[4][1])
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"%s" % ex})

        elif action == 'addsedeexamen':
            try:
                from inno.models import MatriculaSedeExamen
                f = MatriculaSedeExamenForm(request.POST)
                if not 'id' in request.POST:
                    raise NameError(u"Matrícula no encontrada")
                if not Matricula.objects.filter(pk=int(request.POST['id'])).exists():
                    raise NameError(u"Matrícula no encontrada")
                eMatricula = Matricula.objects.get(pk=int(request.POST['id']))

                if not f.is_valid():
                    raise NameError('Formulario incorrecto')

                if MatriculaSedeExamen.objects.values("id").filter(matricula=eMatricula, sede=f.cleaned_data['sede'], detallemodeloevaluativo=f.cleaned_data['detallemodeloevaluativo']).exists():
                    raise NameError(u"El registro ya existe")

                eMatriculaSedeExamen = MatriculaSedeExamen(matricula=eMatricula,
                                                           sede=f.cleaned_data['sede'],
                                                           detallemodeloevaluativo=f.cleaned_data['detallemodeloevaluativo']
                                                           )
                eMatriculaSedeExamen.save(request)
                log(u'Adiciono sede de examen: %s' % eMatriculaSedeExamen, request, "add")
                messages.add_message(request, messages.SUCCESS, f'Se guardo correctamente el registro')
                return JsonResponse({"result": "ok"})

            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos. %s" % ex.__str__()})

        elif action == 'editsedeexamen':
            try:
                from inno.models import MatriculaSedeExamen
                f = MatriculaSedeExamenForm(request.POST)
                if not 'id' in request.POST:
                    raise NameError(u"Registro no encontrado")
                if not MatriculaSedeExamen.objects.filter(pk=int(request.POST['id'])).exists():
                    raise NameError(u"Registro no encontrada")
                eMatriculaSedeExamen = MatriculaSedeExamen.objects.get(pk=int(request.POST['id']))
                eMatricula = eMatriculaSedeExamen.matricula
                if not f.is_valid():
                    raise NameError('Formulario incorrecto')
                if MatriculaSedeExamen.objects.filter(matricula=eMatricula, sede=f.cleaned_data['sede'], detallemodeloevaluativo=f.cleaned_data['detallemodeloevaluativo']).exclude(pk=eMatriculaSedeExamen.pk).exists():
                    raise NameError(u"El registro ya existe")
                eMatriculaSedeExamen.sede=f.cleaned_data['sede']
                eMatriculaSedeExamen.detallemodeloevaluativo=f.cleaned_data['detallemodeloevaluativo']
                eMatriculaSedeExamen.save(request)
                log(u'Edito sede de examen: %s' % eMatriculaSedeExamen, request, "edit")
                messages.add_message(request, messages.SUCCESS, f'Se guardo correctamente el registro')
                return JsonResponse({"result": "ok"})

            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos. %s" % ex.__str__()})

        elif action == 'deletesedeexamen':
            try:
                from inno.models import MatriculaSedeExamen
                if not 'id' in request.POST:
                    raise NameError(u"Registro no encontrado")
                if not MatriculaSedeExamen.objects.filter(pk=int(request.POST['id'])).exists():
                    raise NameError(u"Registro no encontrada")
                eEliminar = eMatriculaSedeExamen = MatriculaSedeExamen.objects.get(pk=int(request.POST['id']))
                eMatriculaSedeExamen.delete()
                log(u'Elimino sede de examen: %s' % eEliminar, request, "del")
                return JsonResponse({"result": "ok", "mensaje": "Se elimino correctamente el registro"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos. %s" % ex.__str__()})

        elif action == 'bloqueomatricula':
            try:
                matricula = Matricula.objects.get(pk=request.POST['id'])
                rubrosvencidos = None
                usermoodle = matricula.inscripcion.persona.usuario.username
                accionretiro = matricula.bloqueomatricula
                matricula.bloqueomatricula = not accionretiro
                matricula.save(request)
                # if matricula.inscripcion.carrera.mi_coordinacion2() == 9:
                cnmoodle = connections['db_moodle_virtual'].cursor()
                # else:
                #     cnmoodle = connections['moodle_db'].cursor()
                if usermoodle and accionretiro:
                    # Consulta en mooc_user
                    sql = """Select id From mooc_user Where suspended=1 and username='%s'""" % (usermoodle)
                    cnmoodle.execute(sql)
                    registro = cnmoodle.fetchall()
                    if registro:
                        sql = """Update mooc_user Set suspended=0 Where username='%s'""" % (usermoodle)
                        cnmoodle.execute(sql)
                log(u'%s: %s' % (u'Desbloquear matricula' if accionretiro else u'Bloquear matricula', matricula), request, "edit")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'obtenerSENESCYT':
            try:
                from soap.consumer.senescyt import Titulos
                if not 'id' in request.POST:
                    raise NameError('No se encontro parametro de inscripción')
                eInscripcion = Inscripcion.objects.get(pk=request.POST['id'])
                ePersona = eInscripcion.persona
                identificacion = ePersona.identificacion()
                eTitulos = Titulos(identificacion)
                mistitulos = eTitulos.consultar()
                eInscripcion.actualiza_estado_matricula()
                return JsonResponse({"result": 'ok', "mensaje": 'Se obtuvo la información correctamente'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": 'bad', "mensaje": f"Ocurrio un error: {ex.__str__()}"})

        elif action == 'actionHorarioExamen':
            try:
                id = request.POST.get('id', None)
                if id is None:
                    raise NameError('No se encontro parametro')
                visiblehorarioexamen = request.POST.get('visiblehorarioexamen', 'ocultar')
                visiblehorarioexamen = visiblehorarioexamen == 'mostrar'
                try:
                    eMatricula = Matricula.objects.get(pk=id)
                except ObjectDoesNotExist:
                    raise NameError('No se encontro matricula')
                MateriaAsignada.objects.filter(matricula=eMatricula).update(visiblehorarioexamen=visiblehorarioexamen)
                eMatricula.delete_cache()
                eInscripcion = eMatricula.inscripcion
                eInscripcion.delete_cache()
                ePersona = eInscripcion.persona
                ePersona.delete_cache()
                return JsonResponse({"result": 'ok', "mensaje": f'Se {"activo" if visiblehorarioexamen else "inactivo"} horario de examen correctamente'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": 'bad', "mensaje": f"Ocurrio un error: {ex.__str__()}"})

        elif action == 'actionResetExamen':
            from inno.models import MatriculaSedeExamen
            try:
                id = request.POST.get('id', None)
                if id is None:
                    raise NameError('No se encontro parametro')
                try:
                    eMatriculaSedeExamen = MatriculaSedeExamen.objects.get(pk=id)
                except ObjectDoesNotExist:
                    raise NameError('No se encontro matricula')
                eMatriculaSedeExamen.archivoidentidad = None
                eMatriculaSedeExamen.archivofoto = None
                eMatriculaSedeExamen.aceptotermino = False
                eMatriculaSedeExamen.fechaaceptotermino = None
                eMatriculaSedeExamen.urltermino = None
                eMatriculaSedeExamen.save(request)
                eMatricula = eMatriculaSedeExamen.matricula
                eMatricula.delete_cache()
                eInscripcion = eMatricula.inscripcion
                eInscripcion.delete_cache()
                ePersona = eInscripcion.persona
                ePersona.delete_cache()
                log(u'Restablecio proceso de archivos y acuerdo de terminos y condiciones de examenes finales: %s' % eMatriculaSedeExamen, request, "add")
                return JsonResponse({"result": 'ok', "mensaje": f'Se restablecio proceso de examenes'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": 'bad', "mensaje": f"Ocurrio un error: {ex.__str__()}"})

        elif action == 'deleteAlumnoPlanificacion':
            try:
                from inno.models import MateriaAsignadaPlanificacionSedeVirtualExamen
                if not 'id' in request.POST:
                    raise NameError(u"No se encontro parametro correcto")
                if not MateriaAsignadaPlanificacionSedeVirtualExamen.objects.values("id").filter(pk=request.POST['id']):
                    raise NameError(u"No se encontro la materia a eliminar")
                eMateriaAsignadaPlanificacionSedeVirtualExamen = eDelete = MateriaAsignadaPlanificacionSedeVirtualExamen.objects.get(
                    pk=request.POST['id'])
                if eMateriaAsignadaPlanificacionSedeVirtualExamen.habilitadoexamen:
                    raise NameError(
                        f"Alumno {eMateriaAsignadaPlanificacionSedeVirtualExamen.materiaasignada.matricula.inscripcion.persona} tiene habilitado el examen, no puede eliminarlo")
                eDelete.delete()
                log(u'Elimino planificación de alumno: %s' % eMateriaAsignadaPlanificacionSedeVirtualExamen, request,
                    "del")
                return JsonResponse({"result": True})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": False, 'message': str(ex)})

        elif action == 'mostrarhorarios':
            try:
                accion = request.POST.get('accion', None)
                try:
                    accion = int(accion)
                except:
                    accion = 0
                titulo = f'Proceso de mostrar horarios de admisión en proceso.'
                datos = {'accion': accion, 'periodo': periodo, 'persona': personasesion, 'es_admision': True, 'es_pregrado': False}
                noti = Notificacion(
                    cuerpo='Se inicializo el proceso de mostar horarios de exámenes de admisión.',
                    titulo=titulo, destinatario=personasesion,
                    url='',
                    prioridad=1, app_label='SGA',
                    fecha_hora_visible=datetime.now() + timedelta(days=1), tipo=2,
                    en_proceso=True)
                noti.save(request)
                actualizar_visible_horario_masivo(request=request, data=datos, notif=noti.pk).start()
                return JsonResponse({"result": 'ok', "mensaje": f'Proceso de mostrar horarios en proceso de admisión.'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": 'bad', "mensaje": f"Ocurrio un error: {ex.__str__()}"})

        elif action == 'limpiarcache':
            try:
                titulo = f'Proceso de limpieza de cache de admisión en proceso.'
                datos = {'periodo': periodo, 'persona': personasesion, 'es_admision': True, 'es_pregrado': False}
                noti = Notificacion(
                    cuerpo='Se inicializo el proceso de limpieza de cache de admisión.',
                    titulo=titulo,
                    destinatario=personasesion,
                    url='',
                    prioridad=1, app_label='SGA',
                    fecha_hora_visible=datetime.now() + timedelta(days=1), tipo=2,
                    en_proceso=True)
                noti.save(request)
                limpiar_cache_masivo(request=request, data=datos, notif=noti.pk).start()
                return JsonResponse({"result": 'ok', "mensaje": f'Proceso de mostrar horarios en proceso de admisión.'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": 'bad', "mensaje": f"Ocurrio un error: {ex.__str__()}"})

        return JsonResponse({"result": "bad", "mensaje": u"Solicitud Incorrecta."})
    else:
        data['title'] = u'Listado de inscripciones'
        persona = request.session['persona']
        if 'action' in request.GET:
            action = request.GET['action']

            if action == 'edit_titulos':
                try:
                    data['title'] = u'Titulos profesionales del Alumno'
                    inscripcion = Inscripcion.objects.get(id=int(request.GET['id']))
                    data['id'] = inscripcion.id
                    data['idp'] = inscripcion.persona.id
                    data['personatitulouniversidad'] = personatitulouniversidad = PersonaTituloUniversidad.objects.filter(persona_id=inscripcion.persona.id)
                    return render(request, "inscripciones_admision/edit_titulos.html", data)
                except Exception as ex:
                    pass

            elif action == 'add_titulos_profesionales':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_inscripciones')
                    data['title'] = u'Adicionar Titulos profesionales al Alumno'
                    data['idp'] = int(request.GET['idp'])
                    data['form'] = PersonaTituloUniversidadForm()
                    return render(request, "inscripciones_admision/add_titulos_profesionales.html", data)
                except Exception as ex:
                    pass

            elif action == 'edit_titulos_profesionales':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_inscripciones')
                    data['title'] = u'Editar Titulos profesionales al Alumno'
                    data['id'] = int(request.GET['id'])
                    data['idp'] = int(request.GET['idp'])
                    data['personatitulouniversidad'] = personatitulouniversidad = PersonaTituloUniversidad.objects.get(pk=request.GET['id'])
                    data['form'] = PersonaTituloUniversidadForm(initial={'codigoregistro':personatitulouniversidad.codigoregistro,
                                                                         'fecharegresado':personatitulouniversidad.fecharegresado,
                                                                         'fecharegistro':personatitulouniversidad.fecharegistro,
                                                                         'fechaacta':personatitulouniversidad.fechaacta,
                                                                         'fechainicio':personatitulouniversidad.fechainicio,
                                                                         'tiponivel':personatitulouniversidad.tiponivel,
                                                                         'tipouniversidad':personatitulouniversidad.tipouniversidad,
                                                                         'universidad':personatitulouniversidad.universidad,
                                                                         'nombrecarrera':personatitulouniversidad.nombrecarrera})
                    return render(request, "inscripciones_admision/edit_titulos_profesionales.html", data)
                except Exception as ex:
                    pass

            elif action == 'delete_titulos_profesionales':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_inscripciones')
                    data['title'] = u'Eliminar Titulos profesionales al Alumno'
                    data['idp'] = int(request.GET['idp'])
                    data['personatitulouniversidad'] = PersonaTituloUniversidad.objects.get(pk=request.GET['id'])
                    return render(request, "inscripciones/delete_titulos_profesionales.html", data)
                except Exception as ex:
                    pass

            elif action == 'cambiomalla':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_inscripciones')
                    data['title'] = u'Cambio de malla'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    puede_modificar_inscripcion(request, inscripcion)
                    form = CambiomallaForm()
                    form.mallas(inscripcion.carrera)
                    data['form'] = form
                    return render(request, "inscripciones_admision/cambiomallaadmin.html", data)
                except Exception as ex:
                    pass

            elif action == 'retirocarrera':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_inscripciones')
                    data['title'] = u'Retiro de Carrera'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    puede_modificar_inscripcion(request, inscripcion)
                    data['form'] = RetiradoMateriaForm()
                    return render(request, "inscripciones/retirocarrera.html", data)
                except Exception as ex:
                    pass

            elif action == 'cambionivel':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_inscripciones')
                    data['title'] = u'Cambio de nivel malla'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    puede_modificar_inscripcion(request, inscripcion)
                    data['form'] = CambionivelmallaForm()
                    return render(request, "inscripciones/cambionivel.html", data)
                except Exception as ex:
                    pass

            elif action == 'cambiocategoria':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_inscripciones')
                    data['title'] = u'Cambio de tipo de inscripción'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    puede_modificar_inscripcion(request, inscripcion)
                    data['form'] = InscripcionTipoInscripcionForm()
                    return render(request, "inscripciones/cambiotipoinscripcion.html", data)
                except Exception as ex:
                    pass

            elif action == 'proyectos':
                try:
                    data['title'] = u'Proyecto de vinculacion'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    data['proyectos'] = inscripcion.participanteproyectovinculacion_set.all().order_by('-proyecto__fin')
                    return render(request, "inscripciones/proyectos.html", data)
                except Exception as ex:
                    pass

            elif action == 'cursos':
                try:
                    data['title'] = u'Cursos y Escuelas Complementarias'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    data['cursos'] = inscripcion.matriculacursoescuelacomplementaria_set.all()
                    return render(request, "inscripciones/cursos.html", data)
                except Exception as ex:
                    pass

            # elif action == 'add':
            #     try:
            #         puede_realizar_accion(request, 'sga.puede_modificar_inscripciones')
            #         data['title'] = u'Nueva inscripción'
            #         form = InscripcionForm()
            #         if 'id' in request.GET:
            #             if PreInscrito.objects.filter(id=request.GET['id'], inscripcion=None).exists():
            #                 preinscrito = PreInscrito.objects.get(id=request.GET['id'])
            #                 data['preinscrito_id'] = preinscrito
            #                 form.initial = {'fecha': datetime.now().date(),
            #                                 'nacimiento': datetime.now().date(),
            #                                 'fecha_ingreso': datetime.now().date(),
            #                                 'nombres': preinscrito.nombres,
            #                                 'apellido1': preinscrito.apellido1,
            #                                 'apellido2': preinscrito.apellido2,
            #                                 'cedula': preinscrito.cedula,
            #                                 'sexo': preinscrito.sexo,
            #                                 'telefono': preinscrito.telefono_celular,
            #                                 'telefono_conv': preinscrito.telefono_domicilio,
            #                                 'telefono_trabajo': preinscrito.telefono_trabajo,
            #                                 'prenivelacion': False if preinscrito.pre else True,
            #                                 'email': preinscrito.email,
            #                                 'observacionespre': preinscrito.institucion.nombre if preinscrito.institucion else '',
            #                                 'colegio': preinscrito.institucion.nombre if preinscrito.institucion else '',
            #                                 'carrera': preinscrito.carrera,
            #                                 'direccion': preinscrito.direccion,
            #                                 'comoseinformo': preinscrito.comoseinformo,
            #                                 'comoseinformootras': preinscrito.comoseinformootro}
            #         form.nuevo()
            #         miscarreras = persona.mis_carreras()
            #         if UTILIZA_GRUPOS_ALUMNOS:
            #             form.con_grupos(Grupo.objects.filter(carrera__in=miscarreras))
            #         else:
            #             form.sin_grupos(miscarreras)
            #         if EMAIL_INSTITUCIONAL_AUTOMATICO:
            #             form.emailautomatico()
            #         data['form'] = form
            #         data['email_institucional_automatico'] = EMAIL_INSTITUCIONAL_AUTOMATICO
            #         data['utiliza_grupos_alumnos'] = UTILIZA_GRUPOS_ALUMNOS
            #         data['dominio'] = EMAIL_DOMAIN
            #         data['preguntas_inscripcion'] = PREGUNTAS_INSCRIPCION
            #         data['correo_obligatorio'] = CORREO_OBLIGATORIO
            #         return render(request, "inscripciones_admision/add.html", data)
            #     except Exception as ex:
            #         pass

            # elif action == 'adicionarotracarrera':
            #     try:
            #         puede_realizar_accion(request, 'sga.puede_modificar_inscripciones')
            #         data['title'] = u'Inscripción de alumno en otra carrera'
            #         data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
            #         documentos = inscripcion.documentos_entregados()
            #         tesdrive = inscripcion.documentos_tesdrive()
            #         form = NuevaInscripcionForm(initial={'fecha': datetime.now().date(),
            #                                              'prenivelacion': documentos.pre,
            #                                              'copiarecord': inscripcion.recordacademico_set.all(),
            #                                              'observacionespre': documentos.observaciones_pre,
            #                                              'titulo': documentos.titulo,
            #                                              'acta': documentos.acta,
            #                                              'cedula2': documentos.cedula,
            #                                              'votacion': documentos.votacion,
            #                                              'actaconv': documentos.actaconv,
            #                                              'partida_nac': documentos.partida_nac,
            #                                              'fotos': documentos.fotos,
            #                                              'licencia': tesdrive.licencia,
            #                                              'record': tesdrive.record,
            #                                              'certificado_tipo_sangre': tesdrive.certificado_tipo_sangre,
            #                                              'prueba_psicosensometrica': tesdrive.prueba_psicosensometrica,
            #                                              'certificado_estudios': tesdrive.certificado_estudios})
            #         miscarreras = persona.mis_carreras()
            #         data['form'] = form
            #         data['dominio'] = EMAIL_DOMAIN
            #         return render(request, "inscripciones_admision/adicionarotracarrera.html", data)
            #     except Exception as ex:
            #         pass

            elif action == 'edit':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_inscripciones')
                    data['title'] = u'Editar inscripción'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    puede_modificar_inscripcion(request, inscripcion)
                    documentos = inscripcion.documentos_entregados()
                    tesdrive = inscripcion.documentos_tesdrive()
                    preguntas_inscripcion = inscripcion.preguntas_inscripcion()
                    perfil = inscripcion.persona.mi_perfil()
                    if inscripcion.unidadeducativa_id is None:
                        data['unid'] = 0
                    else:
                        data['unid'] = inscripcion.unidadeducativa_id
                    form = InscripcionForm(initial={'nombres': inscripcion.persona.nombres,
                                                    'apellido1': inscripcion.persona.apellido1,
                                                    'apellido2': inscripcion.persona.apellido2,
                                                    'cedula': inscripcion.persona.cedula,
                                                    'pasaporte': inscripcion.persona.pasaporte,
                                                    'centroinformacion': inscripcion.centroinformacion,
                                                    'paisnacimiento': inscripcion.persona.paisnacimiento,
                                                    'provincianacimiento': inscripcion.persona.provincianacimiento,
                                                    'cantonnacimiento': inscripcion.persona.cantonnacimiento,
                                                    'parroquianacimiento': inscripcion.persona.parroquianacimiento,
                                                    'nacionalidad': inscripcion.persona.nacionalidad,
                                                    'nacimiento': inscripcion.persona.nacimiento,
                                                    'sexo': inscripcion.persona.sexo,
                                                    'lgtbi': inscripcion.persona.lgtbi,
                                                    'raza': perfil.raza_id,
                                                    'nacionalidadindigena': perfil.nacionalidadindigena_id,
                                                    'sangre': inscripcion.persona.sangre,
                                                    'pais': inscripcion.persona.pais,
                                                    'provincia': inscripcion.persona.provincia,
                                                    'canton': inscripcion.persona.canton,
                                                    'parroquia': inscripcion.persona.parroquia,
                                                    'sector': inscripcion.persona.sector,
                                                    'direccion': inscripcion.persona.direccion,
                                                    'direccion2': inscripcion.persona.direccion2,
                                                    'num_direccion': inscripcion.persona.num_direccion,
                                                    'telefono': inscripcion.persona.telefono,
                                                    'telefono_conv': inscripcion.persona.telefono_conv,
                                                    'email': inscripcion.persona.email.strip(),
                                                    'emailinst': inscripcion.persona.emailinst.strip(),
                                                    'fecha': inscripcion.fecha,
                                                    'grupo': inscripcion.grupo(),
                                                    'sede': inscripcion.sede,
                                                    'carrera': inscripcion.carrera,
                                                    'modalidad': inscripcion.modalidad,
                                                    'sesion': inscripcion.sesion,
                                                    'unidadeducativa': inscripcion.unidadeducativa,
                                                    # 'colegio': inscripcion.colegio,
                                                    'especialidad': inscripcion.especialidad,
                                                    'identificador': inscripcion.identificador,
                                                    'prenivelacion': documentos.pre,
                                                    'observacionespre': documentos.observaciones_pre,
                                                    'comoseinformo': preguntas_inscripcion.comoseinformo,
                                                    'comoseinformootras': preguntas_inscripcion.comoseinformootras,
                                                    'razonesmotivaron': preguntas_inscripcion.razonesmotivaron,
                                                    'titulo': documentos.titulo,
                                                    'acta': documentos.acta,
                                                    'certificado_estudios': tesdrive.certificado_estudios,
                                                    'cedula2': documentos.cedula,
                                                    'votacion': documentos.votacion,
                                                    'partida_nac': documentos.partida_nac,
                                                    'actaconv': documentos.actaconv,
                                                    'fotos': documentos.fotos,
                                                    'licencia': tesdrive.licencia,
                                                    'record': tesdrive.record,
                                                    'certificado_tipo_sangre': tesdrive.certificado_tipo_sangre,
                                                    'prueba_psicosensometrica': tesdrive.prueba_psicosensometrica})
                    form.editar(inscripcion)
                    form.sin_trabajo()
                    miscarreras = persona.mis_carreras()
                    if UTILIZA_GRUPOS_ALUMNOS:
                        form.con_grupos(Grupo.objects.filter(carrera__in=miscarreras))
                    else:
                        form.sin_grupos(miscarreras)
                    data['utiliza_grupos_alumnos'] = UTILIZA_GRUPOS_ALUMNOS
                    data['form'] = form
                    data['email_domain'] = EMAIL_DOMAIN
                    data['email_institucional_automatico'] = EMAIL_INSTITUCIONAL_AUTOMATICO
                    data['preguntas_inscripcion'] = PREGUNTAS_INSCRIPCION
                    data['correo_obligatorio'] = CORREO_OBLIGATORIO
                    return render(request, "inscripciones_admision/edit.html", data)
                except Exception as ex:
                    pass

            elif action == 'extracurricular':
                try:
                    data['title'] = u'Actividades extracurriculares'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    data['pasantias'] = inscripcion.pasantias()
                    data['talleres'] = inscripcion.talleres()
                    data['practicas'] = inscripcion.practicas()
                    data['vccs'] = inscripcion.vcc()
                    return render(request, "inscripciones/extracurricular.html", data)
                except Exception as ex:
                    pass

            elif action == 'record':
                try:
                    data['title'] = u'Registro academico'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    data['records'] = inscripcion.recordacademico_set.filter(status=True).order_by('asignaturamalla__nivelmalla', 'asignatura', 'fecha')
                    data['total_creditos'] = inscripcion.total_creditos()
                    data['total_creditos_malla'] = inscripcion.total_creditos_malla()
                    data['total_creditos_modulos'] = inscripcion.total_creditos_modulos()
                    data['total_creditos_otros'] = inscripcion.total_creditos_otros()
                    data['total_horas'] = inscripcion.total_horas()
                    data['promedio'] = inscripcion.promedio_record()
                    data['aprobadas'] = inscripcion.recordacademico_set.filter(aprobada=True, valida=True).count()
                    data['reprobadas'] = inscripcion.recordacademico_set.filter(aprobada=False, valida=True).count()
                    data['reporte_0'] = obtener_reporte("record_alumno")
                    data['tiene_permiso'] = persona.persona_tiene_permiso(inscripcion.id)
                    # return render(request, "inscripciones/record.html", data)
                    return render(request, "inscripciones_admision/recordadm.html", data)
                except Exception as ex:
                    pass

            elif action == 'editmallahistorica':
                try:
                    data['title'] = u'Añadir malla historica'
                    data['record'] =record= RecordAcademico.objects.get(pk=request.GET['rec'])
                    form = MallaHistoricaForm(initial={"asignaturamallahistorico": record.asignaturamallahistorico,})
                    form.solo_mallas_carrera(record.inscripcion)
                    data['form'] = form
                    return render(request, "inscripciones_admision/editmallahistorica.html", data)
                except Exception as ex:
                    pass


            elif action == 'historico':
                try:
                    data['title'] = u'Historico de notas'
                    data['inscripcion'] = Inscripcion.objects.get(pk=request.GET['id'])
                    data['record'] = record = RecordAcademico.objects.get(pk=request.GET['rec'])
                    data['historicos'] = record.historicorecordacademico_set.all().order_by('-fecha')
                    return render(request, "inscripciones_admision/historico.html", data)
                except Exception as ex:
                    pass

            elif action == 'addrecord':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_records')
                    data['title'] = u'Adicionar registro academico'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    puede_modificar_inscripcion(request, inscripcion)
                    form = RecordAcademicoForm()
                    form.solo_mallas_carrera(inscripcion)
                    data['form'] = form
                    return render(request, "inscripciones/addrecord.html", data)
                except Exception as ex:
                    pass

            elif action == 'addrecordhomologada':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_homologaciones')
                    data['title'] = u'Adicionar homologación'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    puede_modificar_inscripcion(request, inscripcion)
                    form = RecordAcademicoForm()
                    form.homologacion()
                    data['form'] = form
                    return render(request, "inscripciones/addrecordhomologada.html", data)
                except Exception as ex:
                    pass

            elif action == 'addhistorico':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_records')
                    data['title'] = u'Adicionar historico de registro academico'
                    data['record'] = record = RecordAcademico.objects.get(pk=request.GET['idr'])
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    puede_modificar_inscripcion(request, inscripcion)
                    form = HistoricoRecordAcademicoForm(initial={"fecha": datetime.now().date(),
                                                                 "creditos": 0,
                                                                 "horas": 0,
                                                                 "nota": 0,
                                                                 "asistencia": 0})
                    form.solo_asignatura(record.asignatura.id)
                    data['form'] = form
                    return render(request, "inscripciones_admision/addhistorico.html", data)
                except Exception as ex:
                    pass

            elif action == 'matricular':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_matriculas')
                    inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    puede_modificar_inscripcion(request, inscripcion)
                    periodo = request.session['periodo']
                    if MATRICULACION_LIBRE:
                        hoy = datetime.now().date()
                        if Nivel.objects.filter(nivellibrecoordinacion__coordinacion__carrera=inscripcion.carrera, nivellibrecoordinacion__coordinacion__sede=inscripcion.sede, sesion=inscripcion.sesion, modalidad=inscripcion.modalidad, cerrado=False, fin__gte=hoy, periodo=periodo, periodo__tipo__id=TIPO_PERIODO_REGULAR).exists():
                            nivel = Nivel.objects.filter(nivellibrecoordinacion__coordinacion__carrera=inscripcion.carrera, nivellibrecoordinacion__coordinacion__sede=inscripcion.sede, sesion=inscripcion.sesion, modalidad=inscripcion.modalidad, cerrado=False, fin__gte=hoy, periodo=periodo, periodo__tipo__id=TIPO_PERIODO_REGULAR).order_by('-fin')[0]
                            return HttpResponseRedirect('/matriculas_admision?action=addmatriculalibre&id=' + str(nivel.id) + '&iid=' + str(inscripcion.id))
                    return HttpResponseRedirect("/inscripciones_admision?id=" + request.GET['id'])
                except Exception as ex:
                    pass

            elif action == 'edithistorico':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_records')
                    data['title'] = u'Editar historico de registro academico'
                    historico = HistoricoRecordAcademico.objects.get(pk=request.GET['id'])
                    puede_modificar_inscripcion(request, historico.inscripcion)
                    form = HistoricoRecordAcademicoForm(initial={"asignatura": historico.asignatura,
                                                                 "creditos": historico.creditos,
                                                                 "horas": historico.horas,
                                                                 "nota": historico.nota,
                                                                 "asistencia": historico.asistencia,
                                                                 "fecha": historico.fecha,
                                                                 "noaplica": historico.noaplica,
                                                                 "aprobada": historico.aprobada,
                                                                 "valida": historico.valida,
                                                                 "validapromedio": historico.validapromedio,
                                                                 "convalidacion": historico.convalidacion,
                                                                 "homologada": historico.homologada,
                                                                 "observaciones": historico.observaciones})
                    form.editar(historico)
                    data['form'] = form
                    data['historico'] = historico
                    return render(request, "inscripciones_admision/edithistorico.html", data)
                except Exception as ex:
                    pass

            elif action == 'delrecord':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_records')
                    data['title'] = u'Eliminar registro academico'
                    data['record'] = record = RecordAcademico.objects.get(pk=request.GET['id'])
                    puede_modificar_inscripcion(request, record.inscripcion)
                    return render(request, "inscripciones_admision/delrecord.html", data)
                except Exception as ex:
                    pass

            elif action == 'delhistorico':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_records')
                    data['title'] = u'Eliminar historico de registro academico'
                    data['historico'] = historico = HistoricoRecordAcademico.objects.get(pk=request.GET['id'])
                    puede_modificar_inscripcion(request, historico.inscripcion)
                    return render(request, "inscripciones_admision/delhistorico.html", data)
                except Exception as ex:
                    pass

            elif action == 'convalidar':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_homologaciones')
                    data['title'] = u'Homologación de materia'
                    record = RecordAcademico.objects.get(pk=request.GET['id'])
                    puede_modificar_inscripcion(request, record.inscripcion)
                    historico = record.mi_historico()
                    convalidacion = historico.datos_convalidacion()
                    data['form'] = ConvalidacionInscripcionForm(initial={'centro': convalidacion.centro,
                                                                         'carrera': convalidacion.carrera,
                                                                         'asignatura': convalidacion.asignatura,
                                                                         'anno': convalidacion.anno,
                                                                         'nota_ant': convalidacion.nota_ant,
                                                                         'nota_act': convalidacion.nota_act,
                                                                         'creditos': convalidacion.creditos,
                                                                         'observaciones': convalidacion.observaciones})
                    data['convalidacion'] = convalidacion
                    return render(request, "inscripciones/convalidar.html", data)
                except Exception as ex:
                    pass

            elif action == 'homologar':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_homologaciones')
                    data['title'] = u'Homologacion de materia'
                    record = RecordAcademico.objects.get(pk=request.GET['id'])
                    puede_modificar_inscripcion(request, record.inscripcion)
                    historico = record.mi_historico()
                    homologacion = historico.datos_homologacion()
                    data['form'] = HomologacionInscripcionForm(initial={'carrera': homologacion.carrera,
                                                                        'asignatura': homologacion.asignatura,
                                                                        'fecha': homologacion.fecha,
                                                                        'nota_ant': homologacion.nota_ant,
                                                                        'creditos': homologacion.creditos,
                                                                        'modalidad': homologacion.modalidad,
                                                                        'observaciones': homologacion.observaciones})
                    data['record'] = record
                    data['homologacion'] = homologacion
                    return render(request, "inscripciones/homologar.html", data)
                except Exception as ex:
                    transaction.set_rollback(True)
                    pass

            elif action == 'cargarfoto':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_inscripciones')
                    data['title'] = u'Subir foto'
                    data['form'] = CargarFotoForm()
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    puede_modificar_inscripcion(request, inscripcion)
                    return render(request, "inscripciones/cargarfoto.html", data)
                except Exception as ex:
                    pass

            elif action == 'documentos':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_inscripciones')
                    data['title'] = u'Documentos y archivos'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    puede_modificar_inscripcion(request, inscripcion)
                    data['archivos'] = inscripcion.archivo_set.all().order_by('fecha')
                    data['docperfil'] = inscripcion.persona.mi_perfil()
                    data['docreligion'] = PersonaReligion.objects.filter(status=True, persona= inscripcion.persona)
                    data['docestudiante'] = inscripcion.persona
                    data['docpersonal'] = PersonaDocumentoPersonal.objects.filter(status=True, persona_id=inscripcion.persona_id).first()
                    #PersonaDocumentoPersonal.objects.get(status=True, persona_id=inscripcion.persona.usuario_id)


                    return render(request, "inscripciones_admision/documentos.html", data)
                except Exception as ex:
                    pass

            elif action == 'adddocumento':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_inscripciones')
                    data['title'] = u'Adicionar archivos'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    puede_modificar_inscripcion(request, inscripcion)
                    data['form'] = DocumentoInscripcionForm()
                    return render(request, "inscripciones_admision/adddocumento.html", data)
                except Exception as ex:
                    pass

            elif action == 'importar':
                try:
                    puede_realizar_accion(request, 'sga.puede_importar_inscripciones')
                    data['title'] = u'Importar inscripciones'
                    data['form'] = ImportarArchivoXLSForm()
                    return render(request, "inscripciones/importar.html", data)
                except Exception as ex:
                    pass

            elif action == 'deldocumento':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_inscripciones')
                    data['title'] = u'Eliminar archivo o documento'
                    data['archivo'] = archivo = Archivo.objects.get(pk=request.GET['id'])
                    puede_modificar_inscripcion(request, archivo.inscripcion)
                    return render(request, "inscripciones_admision/deldocumento.html", data)
                except Exception as ex:
                    pass

            elif action == 'fechainicioconvalidacion':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_inscripciones')
                    data['title'] = u'Fecha inicio convalidacion'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    puede_modificar_inscripcion(request, inscripcion)
                    data['form'] = FechaInicioConvalidacionInscripcionForm(initial={'fecha': inscripcion.fechainicioconvalidacion if inscripcion.fechainicioconvalidacion else datetime.now().date()})
                    return render(request, "inscripciones/fechainicioconvalidacion.html", data)
                except Exception as ex:
                    pass

            elif action == 'fechainicioprimernivel':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_inscripciones')
                    data['title'] = u'Fecha inicio primer nivel'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    puede_modificar_inscripcion(request, inscripcion)
                    data['form'] = FechaInicioPrimerNivelInscripcionForm(initial={'fecha': inscripcion.fechainicioprimernivel if inscripcion.fechainicioprimernivel else datetime.now().date()})
                    return render(request, "inscripciones/fechainicioprimernivel.html", data)
                except Exception as ex:
                    pass

            elif action == 'alumalla':
                try:
                    data['title'] = u'Malla del alumno'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    data['inscripcion_malla'] = inscripcionmalla = inscripcion.malla_inscripcion()
                    data['malla'] = malla = inscripcionmalla.malla
                    data['nivelesdemallas'] = NivelMalla.objects.all().order_by('id')
                    data['ejesformativos'] = EjeFormativo.objects.all().order_by('nombre')
                    data['asignaturasmallas'] = [(x, inscripcion.aprobadaasignatura(x)) for x in AsignaturaMalla.objects.filter(malla=malla)]
                    resumenniveles = [{'id': x.id, 'horas': x.total_horas(malla), 'creditos': x.total_creditos(malla)} for x in NivelMalla.objects.all().order_by('id')]
                    data['resumenes'] = resumenniveles
                    return render(request, "inscripciones_admision/alumalla.html", data)
                except Exception as ex:
                    pass

            elif action == 'alucarrera':
                try:
                    data['title'] = u'Carreras del alumno'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    data['persona'] = alumno = inscripcion.persona
                    data['inscripciones'] = alumno.inscripcion_set.filter(status=True).exclude(carrera__coordinacion__id=9)
                    return render(request, "inscripciones_admision/alucarrera.html", data)
                except Exception as ex:
                    pass

            elif action == 'seguimiento_asignaturas_alumno':
                try:
                    porcentaje_general = []
                    porcentaje_ponderacion = []
                    data['title'] = u'Seguimiento Asignaturas'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    data['materiassga'] = materiassga = inscripcion.materias(periodo)
                    if inscripcion.matricula_admision_virtual(periodo):
                        data['modeloevaluativo'] = inscripcion.matricula_admision_virtual(periodo).mis_materias_sin_retiro()[0].materia.modeloevaluativo
                    else:
                        data['modeloevaluativo'] = None
                    for materia in materiassga:
                        porcentaje_general.append([materia.materia.asignatura.nombre,str(inscripcion.promedio_general_por_asignatura_migradas1(materia.materia.id))])
                        # porcentaje_ponderacion.append([materia.materia.asignatura.nombre,str(inscripcion.porcentaje_por_asignatura_con_examen_migradas(materia.materia.id))])
                    data['porcentaje_general'] = porcentaje_general
                    # data['porcentaje_ponderacion'] = porcentaje_ponderacion
                    return render(request, "inscripciones_admision/actividadessakai.html", data)
                except Exception as ex:
                    pass

            elif action == 'actividadesmoodle':
                try:
                    data['title'] = u'Actividades sistema Moodle'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    data['materiassga'] = inscripcion.materias(periodo)
                    return render(request, "inscripciones_admision/actividadesmoodle.html", data)
                except Exception as ex:
                    pass


            elif action == 'porcentajeactividadgeneral':
                try:
                    data['title'] = u'Porcentaje de Actividades Generales Sistema SAKAI'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    if periodo.usa_sakai:
                        data['id_usuario'] = id_usuario = inscripcion.persona.codigo_sakai()
                        data['resultados_cursos'] = inscripcion.asignaturas_sakai()
                        return render(request, "inscripciones_admision/porcentajeactividadgeneral.html", data)
                    else:
                        data['materiassga'] = inscripcion.materias(periodo)
                        return render(request, "inscripciones_admision/porcentajeactividadgeneral1.html", data)
                except Exception as ex:
                    pass

            elif action == 'actividadesalumno':
                try:
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['idinscripcion'])
                    if periodo.usa_sakai:
                        data['title'] = u'Listado de tareas'
                        data['idcurso'] = idcurso = request.GET['idcurso']
                        data['nomcurso'] = request.GET['nomcurso']
                        data['id_usuario'] = id_usuario = inscripcion.persona.codigo_sakai()
                        data['listas_tareas'] = inscripcion.tareas_por_asignatura_con_datos(idcurso)
                        data['resultados_foros'] = resultados_foros = inscripcion.lista_foros_puntos_posibles_calificado(idcurso)
                        data['resultados_chats'] = resultados_chats = inscripcion.lista_chats_asignatura(idcurso)
                        # MENSAJES ENVIADOS
                        data['num_mensajes_enviados'] = resultados_mensajes_enviados = inscripcion.numero_mensajes_enviados_asignatura(idcurso)
                        data['num_mensajes_recibidos_sin_leer'] = num_mensajes_recibidos_sin_leer = inscripcion.numero_mensajes_recibidos_sin_leer(idcurso)
                        data['num_mensajes_recibidos_leidos'] = num_mensajes_recibidos_leidos = inscripcion.numero_mensajes_recibidos_leidos(idcurso)
                        # EXTRAER EXAMENES
                        data['datosexamenes']= inscripcion.lista_test_asignatura_estudiante(idcurso)
                        # data['modeloevaluativo'] = inscripcion.matricula().mis_materias_sin_retiro()[0].materia.modeloevaluativo
                        return render(request, "inscripciones_admision/listatareas.html", data)
                    elif periodo.usa_moodle:
                        data['coord'] = inscripcion.carrera.coordinacion_set.all()[0].id
                        data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['idinscripcion'])
                        data['materia'] = materia = Materia.objects.get(id= request.GET['idcurso'])
                        return render(request, "inscripciones_admision/actividadesalumno.html", data)
                except Exception as ex:
                    pass

            elif action == 'porcentajeactividad':
                try:
                    data['title'] = u'Porcentaje de Actividades'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['idinscripcion'])
                    if periodo.usa_sakai:
                        data['idcurso'] = idcurso = request.GET['idcurso']
                        data['nomcurso'] = request.GET['nomcurso']
                        id_usuario = inscripcion.persona.codigo_sakai()
                        data['id_usuario'] = id_usuario
                        return render(request, "inscripciones_admision/porcentajeactividad.html", data)
                    elif periodo.usa_moodle:
                        data['materia'] = Materia.objects.filter(id=int(request.GET['idcurso']))
                        return render(request, "inscripciones_admision/porcentajeactividadmigrada.html", data)
                except Exception as ex:
                    pass

            elif action == 'novalidar':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_records')
                    data['title'] = u'No considerar créditos'
                    data['record'] = record = RecordAcademico.objects.get(pk=request.GET['id'])
                    puede_modificar_inscripcion(request, record.inscripcion)
                    data['form'] = ConsiderarForm()
                    return render(request, "inscripciones_admision/novalidar.html", data)
                except Exception as ex:
                    pass

            elif action == 'validar':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_records')
                    data['title'] = u'Considerar créditos'
                    data['record'] = record = RecordAcademico.objects.get(pk=request.GET['id'])
                    puede_modificar_inscripcion(request, record.inscripcion)
                    data['form'] = ConsiderarForm()
                    return render(request, "inscripciones_admision/validar.html", data)
                except Exception as ex:
                    pass

            elif action == 'novalidarpromedio':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_records')
                    data['title'] = u'No considerar para promedio'
                    data['record'] = record = RecordAcademico.objects.get(pk=request.GET['id'])
                    puede_modificar_inscripcion(request, record.inscripcion)
                    data['form'] = ConsiderarForm()
                    return render(request, "inscripciones_admision/novalidarpromedio.html", data)
                except Exception as ex:
                    pass

            elif action == 'validarpromedio':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_records')
                    data['title'] = u'Considerar para promedio'
                    data['record'] = record = RecordAcademico.objects.get(pk=request.GET['id'])
                    puede_modificar_inscripcion(request, record.inscripcion)
                    data['form'] = ConsiderarForm()
                    return render(request, "inscripciones_admision/validarpromedio.html", data)
                except Exception as ex:
                    pass

            elif action == 'cambiogrupo':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_inscripciones')
                    data['title'] = u'Cambio de grupo'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    puede_modificar_inscripcion(request, inscripcion)
                    form = CambioGrupoForm()
                    form.cambio_grupo(inscripcion)
                    data['form'] = form

                    return render(request, "inscripciones_admision/cambiogrupo.html", data)
                except Exception as ex:
                    pass

            elif action == 'recalcularcreditos':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_records')
                    inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    puede_modificar_inscripcion(request, inscripcion)
                    inscripcion.actualizar_creditos()
                    return HttpResponseRedirect("/inscripciones_admision?action=record&id=" + request.GET['id'])
                except Exception as ex:
                    transaction.set_rollback(True)
                    pass

            elif action == 'recalcularniveles':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_records')
                    inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    puede_modificar_inscripcion(request, inscripcion)
                    inscripcion.actualizar_niveles_records()
                    inscripcion.actualizar_nivel()
                    return HttpResponseRedirect("/inscripciones_admision?action=record&id=" + request.GET['id'])
                except Exception as ex:
                    transaction.set_rollback(True)
                    pass

            elif action == 'desactivar':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_inscripciones')
                    data['title'] = u'Desactivar usuario'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    puede_modificar_inscripcion(request, inscripcion)
                    return render(request, "inscripciones_admision/desactivar.html", data)
                except Exception as ex:
                    pass

            elif action == 'activar':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_inscripciones')
                    data['title'] = u'Activar usuario'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    puede_modificar_inscripcion(request, inscripcion)
                    return render(request, "inscripciones_admision/activar.html", data)
                except Exception as ex:
                    pass

            elif action == 'desactivarperfil':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_administrativos')
                    data['title'] = u'Desactivar perfil de usuario'
                    data['inscripcion'] = Inscripcion.objects.get(pk=request.GET['id'])
                    return render(request, "inscripciones_admision/desactivarperfil.html", data)
                except Exception as ex:
                    pass

            elif action == 'activarperfil':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_administrativos')
                    data['title'] = u'Activar perfil de usuario'
                    data['inscripcion'] = Inscripcion.objects.get(pk=request.GET['id'])
                    return render(request, "inscripciones_admision/activarperfil.html", data)
                except Exception as ex:
                    pass

            elif action == 'resetear':
                try:
                    # puede_realizar_accion(request, 'sga.puede_modificar_inscripciones')
                    data['title'] = u'Resetear clave del usuario'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    # puede_modificar_inscripcion(request, inscripcion)
                    return render(request, "inscripciones_admision/resetear.html", data)
                except Exception as ex:
                    pass

            elif action == 'resetear_clave_admision_virtual':
                try:
                    data['title'] = u'Resetear clave del usuario'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    return render(request, "inscripciones_admision/resetear_clave_admision_virtual.html", data)
                except Exception as ex:
                    pass

            elif action == 'asignar_estudiantes_admin':
                try:
                    data['title'] = u'Asignar todos matriculados del periodo'
                    data['soporte'] = soporte =  VirtualSoporteUsuario.objects.get(id=request.GET['ids'])
                    return render(request, "inscripciones_admision/asignar_estudiantes_admin.html", data)
                except Exception as ex:
                    pass

            elif action == 'asignarestudiantescarrera':
                try:
                    from itertools import chain
                    data['title'] = u'Asignar Estudiantes'
                    data['soporte'] = soporte = VirtualSoporteUsuario.objects.get(id=request.GET['ids'])
                    carreras = soporte.virtualsoporteusuarioinscripcion_set.values_list('matricula__inscripcion__carrera__id',flat=True).filter(status=True, matricula__nivel__periodo=periodo).distinct()
                    carreras2id = Matricula.objects.values_list('inscripcion__carrera__id',flat=True).filter(status=True, nivel__periodo=periodo).distinct()
                    ids = list(chain(carreras, carreras2id))
                    form = AsignarSoporteEstudiante(initial={'carrera': Carrera.objects.filter(status=True, pk__in=carreras)})
                    form.cargar_carreras(ids)
                    data['form'] = form
                    return render(request, "niveles/asignar_estudiantes_carrera.html", data)
                except Exception as ex:
                    pass


            elif action == 'actividades':
                try:
                    data['title'] = u'Actividades extracurriculares'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    data['actividades'] = inscripcion.participanteactividadextracurricular_set.all().order_by('-actividad__fechafin')
                    return render(request, "inscripciones_admision/actividades.html", data)
                except Exception as ex:
                    pass

            elif action == 'horario':
                try:
                    data['title'] = u'Horario del estudiante'
                    data['matricula'] = matricula = Matricula.objects.get(pk=request.GET['id'])
                    data['semana'] = [[1, 'Lunes'], [2, 'Martes'], [3, 'Miercoles'], [4, 'Jueves'], [5, 'Viernes'], [6, 'Sabado'], [7, 'Domingo']]
                    hoy = datetime.now().date()
                    data['misclases'] = clases = Clase.objects.filter(activo=True, fin__gte=hoy, materia__materiaasignada__matricula=matricula).order_by('inicio')
                    data['inscripcion'] = matricula.inscripcion
                    data['sesiones'] = Sesion.objects.filter(turno__clase__in=clases).distinct()
                    return render(request, "inscripciones_admision/horario.html", data)
                except Exception as ex:
                    pass

            elif action == 'addadministrativo':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_administrativos')
                    data['title'] = u'Crear perfil de administrativo'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    return render(request, "inscripciones_admision/addadministrativo.html", data)
                except Exception as ex:
                    pass

            elif action == 'adddocente':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_profesores')
                    data['title'] = u'Crear perfil de profesor'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    return render(request, "inscripciones_admision/adddocente.html", data)
                except Exception as ex:
                    pass

            elif action == 'seguimiento_estudiantes':
                try:
                    data['title'] = u'Seguimiento de Actividades del estudiante'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    data['materiassga'] = materiassga = inscripcion.materias(periodo)
                    return render(request, "inscripciones_admision/seguimiento_estudiantes.html", data)
                except Exception as ex:
                    pass

            elif action == 'listadosonline':
                try:
                    data['title'] = u'Listado de Inscritos online'
                    search = None
                    ids = None
                    tipobus = None
                    inscripcionid = None
                    carr = None
                    listapersonainscripcion = Inscripcion.objects.filter(carrera__coordinacion__id=9).order_by('persona__apellido1','persona__apellido2')
                    if 'id' in request.GET:
                        ids = request.GET['id']
                        listapersonainscripcion = listapersonainscripcion.filter(id=ids)
                    elif 's' in request.GET:
                        search = request.GET['s'].strip()
                        ss = search.split(' ')
                        if len(ss) == 1:
                            listapersonainscripcion = listapersonainscripcion.filter(Q(persona__nombres__icontains=search) |
                                                                                     Q(persona__apellido1__icontains=search) |
                                                                                     Q(persona__apellido2__icontains=search) |
                                                                                     Q(persona__cedula__icontains=search) |
                                                                                     Q(persona__pasaporte__icontains=search) |
                                                                                     Q(identificador__icontains=search) |
                                                                                     Q(inscripciongrupo__grupo__nombre__icontains=search) |
                                                                                     Q(persona__usuario__username__icontains=search)).distinct()
                        else:
                            listapersonainscripcion = listapersonainscripcion.filter(Q(persona__apellido1__icontains=ss[0]) &
                                                                                     Q(persona__apellido2__icontains=ss[1])).distinct()
                    elif 'carr' in request.GET:
                        carr=int(request.GET['carr'])
                        if carr>0:
                            listapersonainscripcion = listapersonainscripcion.filter(carrera__id=carr)
                    paging = MiPaginador(listapersonainscripcion, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['page'] = page
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['totalonline'] = listapersonainscripcion.filter(modalidad_id=3, confimacion_online=True).count()
                    data['carreras'] = Carrera.objects.filter(id__in=Matricula.objects.values_list('inscripcion__carrera__id', flat=True).filter(nivel__periodo=periodo, inscripcion__carrera__coordinacion__id=9)).order_by('modalidad')
                    data['cant']=listapersonainscripcion.count()
                    data['pais'] = Pais.objects.filter(id__in=listapersonainscripcion.values_list('persona__pais__id',flat=True).filter(persona__pais__status=True))
                    data['totalemailenviados'] = listapersonainscripcion.filter(modalidad_id=3, envioemail=True).count()
                    data['listapersonainscripcion'] = page.object_list
                    data['search'] = search if search else ""
                    data['carreraselect'] = carr if carr else ""
                    data['ids'] = ids if ids else ""
                    return render(request, "inscripciones_admision/listadosonline.html", data)
                except Exception as ex:
                    pass

            elif action == 'imprimiractividadalumno':
                try:
                    inscripcion = Inscripcion.objects.get(pk=int(str(request.GET['idinscripcion'])))
                    idcurso = str(request.GET['idcurso'])
                    nomcurso = str(request.GET['nomcurso'])
                    if periodo.usa_sakai:
                        listas_tareas=[]
                        id_usuario = inscripcion.persona.codigo_sakai()
                        listas_tareas= inscripcion.tareas_por_asignatura_con_datos(idcurso)
                        resultados_foros = inscripcion.lista_foros_puntos_posibles_calificado(idcurso)

                        resultados_chats = inscripcion.lista_chats_asignatura(idcurso)
                        # MENSAJES ENVIADOS
                        num_mensajes_enviados = inscripcion.numero_mensajes_enviados_asignatura(idcurso)
                        # EXTRAER MENSAJES RECIBIDOS
                        num_mensajes_recibidos_sin_leer = inscripcion.numero_mensajes_recibidos_sin_leer(idcurso)
                        # EXTRAER MENSAJES RECIBIDOS LEIDOS
                        num_mensajes_recibidos_leidos = inscripcion.numero_mensajes_recibidos_leidos(idcurso)
                        # EXTRAER EXAMENES
                        datosexamenes = inscripcion.lista_test_asignatura_estudiante(idcurso)
                        return conviert_html_to_pdf('inscripciones_admision/imprimiractividadalumno.html',
                                                    {'pagesize': 'A4',
                                                     'data': data,'idcurso':idcurso,
                                                     'inscripcion':inscripcion,
                                                     'nomcurso':nomcurso,'listas_tareas':listas_tareas,'resultados_foros':resultados_foros,'resultados_chats':resultados_chats,
                                                     'num_mensajes_enviados':num_mensajes_enviados,'num_mensajes_recibidos_sin_leer':num_mensajes_recibidos_sin_leer,
                                                     'num_mensajes_recibidos_leidos':num_mensajes_recibidos_leidos, 'datosexamenes':datosexamenes
                                                     })
                    elif periodo.usa_moodle:
                        coord=inscripcion.carrera.coordinacion_set.all()[0].id
                        return conviert_html_to_pdf('inscripciones_admision/imprimiractividadalumnomigrada.html',
                                                    {'pagesize': 'A4',
                                                     'data': data, 'materia': Materia.objects.get(id=idcurso),
                                                     'inscripcion': inscripcion,'coord':coord
                                                     })
                except Exception as ex:
                    pass

            elif action == 'imprimirporcentajeactividad':
                try:
                    data['title'] = u'Porcentaje de Actividades por Asignatura'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['idinscripcion'])
                    data['idcurso'] = idcurso = request.GET['idcurso']
                    data['nomcurso'] = nomcurso= request.GET['nomcurso']
                    data['id_usuario'] = id_usuario = inscripcion.persona.codigo_sakai()

                    numero_tareas_por_asignatura =inscripcion.numero_tareas_por_asignatura(idcurso)
                    numero_tareas_por_asignatura_enviadas =inscripcion.numero_tareas_por_asignatura_enviadas(idcurso)
                    porcentaje_cumplimiento_tareas_asignatura =inscripcion.porcentaje_cumplimiento_tareas_asignatura(idcurso)
                    numero_foro_por_asignatura =inscripcion.numero_foro_por_asignatura(idcurso)
                    numero_foro_a_calificar_asignatura =inscripcion.numero_foro_a_calificar_asignatura(idcurso)
                    numero_foro_participado =inscripcion.numero_foro_participado(idcurso)
                    calcular_porcentaje_foro =inscripcion.calcular_porcentaje_foro(idcurso)
                    numero_test_por_asignatura =inscripcion.numero_test_por_asignatura(idcurso)
                    numero_test_resuelto_por_asignatura =inscripcion.numero_test_resuelto_por_asignatura(idcurso)
                    calcular_porcentaje_test =inscripcion.calcular_porcentaje_test(idcurso)

                    return conviert_html_to_pdf('inscripciones_admision/imprimirporcentajealumno.html',
                                                {'pagesize': 'A4',
                                                 'data': data,
                                                 'inscripcion':inscripcion,
                                                 'nomcurso': nomcurso,
                                                 'numero_tareas_por_asignatura':numero_tareas_por_asignatura,
                                                 'numero_tareas_por_asignatura_enviadas':numero_tareas_por_asignatura_enviadas,
                                                 'porcentaje_cumplimiento_tareas_asignatura':porcentaje_cumplimiento_tareas_asignatura,
                                                 'numero_foro_por_asignatura':numero_foro_por_asignatura,
                                                 'numero_foro_a_calificar_asignatura':numero_foro_a_calificar_asignatura,
                                                 'numero_foro_participado':numero_foro_participado,
                                                 'calcular_porcentaje_foro':calcular_porcentaje_foro,
                                                 'numero_test_por_asignatura':numero_test_por_asignatura,
                                                 'numero_test_resuelto_por_asignatura':numero_test_resuelto_por_asignatura,
                                                 'calcular_porcentaje_test':calcular_porcentaje_test
                                                 })
                except Exception as ex:
                    pass

            elif action == 'imprimirporcentajegeneralalumno':
                try:
                    data['title'] = u'Porcentaje General'
                    inscripcion = Inscripcion.objects.get(pk=int(str(request.GET['idinscripcion'])))
                    data['resultados_cursos'] = resultados_cursos = inscripcion.asignaturas_sakai()
                    return conviert_html_to_pdf('imprimirporcentajeactividadalumnogeneralsakai.html',
                                                {'pagesize': 'A4',
                                                 'data': data,
                                                 'inscripcion':inscripcion,
                                                 'resultados_cursos':resultados_cursos
                                                 })
                except Exception as ex:
                    pass

            elif action == 'verificarestadomatricula':
                try:
                    data['title'] = u'Importar inscripciones'
                    data['form'] = ImportarArchivoXLSForm()
                    return render(request, "inscripciones_admision/verificarestadomatricula.html", data)
                except Exception as ex:
                    pass

            elif action == 'imprimiractividadgeneralalumno':
                try:
                    inscripcion = Inscripcion.objects.get(pk=int(str(request.GET['idinscripcion'])))
                    lista_asignaturas = inscripcion.asignaturas_sakai()
                    return conviert_html_to_pdf('imprimiractividadalumnogeneralsakai.html',
                                                {'pagesize': 'A4',
                                                 'data': data,
                                                 'inscripcion':inscripcion,
                                                 'lista_asignaturas':lista_asignaturas
                                                 })
                except Exception as ex:
                    pass

            elif action == 'resumen':
                try:
                    data['title'] = u'Resumen '
                    return render(request, "inscripciones_admision/resumen.html", data)
                except Exception as ex:
                    pass

            elif action == 'bloqueomatricula':
                try:
                    puede_realizar_accion(request, 'sga.puede_bloquear_matricula')
                    data['matricula'] = mnatricula = Matricula.objects.get(pk=request.GET['id'])
                    data['inscripcion'] = Inscripcion.objects.get(pk=mnatricula.inscripcion.id)
                    data['title'] = u'Desbloquear matricula' if mnatricula.bloqueomatricula else u'Bloquear matricula'
                    return render(request, "inscripciones_admision/bloqueomatricula.html", data)
                except Exception as ex:
                    pass

            elif action == 'listaasignatura':
                try:
                    if 'idcarrera' in request.GET:
                        data['idcarrera'] = idcarrera = request.GET['idcarrera']
                        carrera = Carrera.objects.get(id=idcarrera)
                        lista=[]
                        materias = Materia.objects.filter(status=True,asignaturamalla__malla__carrera=carrera).order_by('asignatura')
                        for x in materias:
                            lista.append([x.id, x.flexbox_repr()])
                        data = {"result": "ok", "lista": lista}
                        return JsonResponse(data)
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            elif action == 'materias_asignadas':
                try:
                    if 'idinscripcion' in request.GET:
                        data['idinscripcion'] = idinscripcion = encrypt(request.GET['idinscripcion'])
                        inscripcion = Inscripcion.objects.get(pk=int(idinscripcion))
                        lista=[]
                        for x in inscripcion.matricula().mis_materias_sin_retiro():
                            lista.append([x.materia.id, x.materia.asignatura.nombre])
                        data = {"result": "ok", "lista": lista}
                        return JsonResponse(data)
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            elif action == 'listar_horas':
                try:
                    if 'dia' in request.GET:
                        dia = request.GET['dia']
                    if 'sede' in request.GET:
                        sede = request.GET['sede']
                    sql = "select distinct h.horario from  horario h where h.dia='"+dia+"' AND h.sede='"+sede+"'"
                    lista = querymysqlconsulta(sql, True)
                    data = {"result": "ok", "lista": lista}
                    return JsonResponse(data)
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            elif action == 'listar_dias':
                try:
                    if 'sede' in request.GET:
                        sede = request.GET['sede']
                    sql = "SELECT DISTINCT h.dia FROM horario h where h.sede = '"+sede+"'"
                    lista = querymysqlconsulta(sql, True)
                    data = {"result": "ok", "lista": lista}
                    return JsonResponse(data)
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            elif action == 'listar_bloque':
                try:
                    if 'sede' in request.GET:
                        sede = request.GET['sede']
                    sql = "SELECT DISTINCT h.bloque FROM horario h where h.sede = '"+sede+"'"
                    lista = querymysqlconsulta(sql, True)
                    data = {"result": "ok", "lista": lista}
                    return JsonResponse(data)
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            elif action == 'listar_labotarios':
                try:
                    if 'bloque' in request.GET:
                        bloque = request.GET['bloque']
                    sql = "select distinct h.laboratorio from  horario h where h.bloque='"+bloque+"'"
                    lista = querymysqlconsulta(sql, True)
                    data = {"result": "ok", "lista": lista}
                    return JsonResponse(data)
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})


            elif action == 'resultados':
                try:
                    idcarrera = 0
                    idmateria = 0
                    uno = 0
                    dos = 0
                    tres = 0
                    cuatro = 0
                    cinco = 0
                    seis = 0
                    siete = 0
                    carrera = None
                    materia = None
                    matriculas = None
                    nombre = ''
                    listado = []
                    if 'idcarrera' in request.GET:
                        idcarrera = int(request.GET['idcarrera'])
                    if 'idmateria' in request.GET:
                        idmateria = int(request.GET['idmateria'])
                    matriculas = Matricula.objects.filter(status=True)
                    if idcarrera > 0:
                        carrera = Carrera.objects.get(id=idcarrera)
                        matriculas.filter(inscripcion__carrera=carrera, nivel__periodo=periodo).distinct()
                        nombre = str(carrera.nombre_completo())
                    if idmateria > 0:
                        materia = Materia.objects.get(id=idmateria)
                        lista = MateriaAsignada.objects.values_list('matricula__id', flat=True).filter(status=True, materia=materia)
                        matriculas = matriculas.filter(status=True, inscripcion__carrera=carrera, nivel__periodo=periodo, id__in=lista).distinct()
                        nombre = nombre + " / " + str(materia.nombre_completo())
                    for matricula in matriculas:
                        porcentaje = 0
                        try:
                            if idmateria > 0:
                                porcentaje = matricula.inscripcion.porcentaje_actividades_por_asignatura(materia.codigosakai)
                            elif idcarrera > 0:
                                porcentaje = matricula.inscripcion.porcentaje_total_estudiante()
                        except Exception as ex:
                            porcentaje = 0

                        if porcentaje <= 10:
                            uno += 1
                        elif porcentaje >= 10.01 and porcentaje <= 20:
                            dos += 1
                        elif porcentaje >= 20.01 and porcentaje <= 30:
                            tres += 1
                        elif porcentaje >= 30.01 and porcentaje <= 40:
                            cuatro += 1
                        elif porcentaje >= 40.01 and porcentaje <= 50:
                            cinco += 1
                        elif porcentaje >= 50.01 and porcentaje <= 60:
                            seis += 1
                        elif porcentaje >= 60.01 and porcentaje <= 70:
                            siete += 1
                    data['uno'] = uno
                    lista = {}
                    lista['nombre'] = '0-10'
                    lista['porcentaje'] = uno
                    listado.append(lista)
                    data['dos'] = dos
                    lista = {}
                    lista['nombre'] = '10.01-20'
                    lista['porcentaje'] = dos
                    listado.append(lista)
                    data['tres'] = tres
                    lista = {}
                    lista['nombre'] = '20.01-30'
                    lista['porcentaje'] = tres
                    listado.append(lista)
                    data['cuatro'] = cuatro
                    lista = {}
                    lista['nombre'] = '30.01-40'
                    lista['porcentaje'] = cuatro
                    listado.append(lista)
                    data['cinco'] = cinco
                    lista = {}
                    lista['nombre'] = '40.01-50'
                    lista['porcentaje'] = cinco
                    listado.append(lista)
                    data['seis'] = seis
                    lista = {}
                    lista['nombre'] = '50.01-60'
                    lista['porcentaje'] = seis
                    listado.append(lista)
                    data['siete'] = siete
                    lista = {}
                    lista['nombre'] = '60.01-70'
                    lista['porcentaje'] = siete
                    listado.append(lista)
                    data['nombre'] = nombre
                    lista_porcentaje = []
                    rango1 = 0
                    rango2 = 0
                    rango3 = 0
                    con1 = 0
                    con2 = 0
                    con3 = 0

                    rango1 = materia.rango_porcentaje()[0]
                    rango2 = materia.rango_porcentaje()[1]
                    if materia.rango_porcentaje().__len__() == 3:
                        rango3 = materia.rango_porcentaje()[2]

                    for matricula in matriculas:
                        porcentaje = matricula.inscripcion.porcentaje_por_asignatura(materia.codigosakai)
                        if porcentaje <= 0:
                            con1 += 1
                        elif porcentaje >= 0 and porcentaje <= rango1:
                            con1 += 1
                        elif porcentaje >= (rango1 + 0.01) and porcentaje <= rango2:
                            con2 += 1
                        elif rango3 > 0:
                            if porcentaje >= (rango2 + 0.01) and porcentaje <= rango3:
                                con3 += 1

                    lista_porcentaje.append({"nombre": str(0) + '% - ' + str(rango1) + '%', "porcentaje": con1})
                    lista_porcentaje.append(
                        {"nombre": str(rango1 + 0.01) + '% - ' + str(rango2) + '%', "porcentaje": con2})
                    if rango3 > 0:
                        lista_porcentaje.append(
                            {"nombre": str(rango2 + 0.01) + '% - ' + str(rango3) + '%', "porcentaje": con3})

                    data['lista_porcentaje'] = lista_porcentaje
                    template = get_template("inscripciones_admision/resultados.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'html': json_content, 'nombre': nombre, 'lista': listado,'lista_porcentaje':lista_porcentaje,'total': matriculas.__len__()})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            elif action == 'listado_alumnos_aprobados_virtual':
                try:
                    __author__ = 'Unemi'
                    matriculas = Matricula.objects.filter(status=True,nivel__periodo=periodo, inscripcion__modalidad__id=3, inscripcion__carrera__coordinacion__id=9).distinct().order_by( 'inscripcion__persona__apellido1', 'inscripcion__persona__apellido2','inscripcion__persona__nombres')
                    title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('alumnos_admision_online')
                    ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=listado_alumnos_aprobados_virtual' + random.randint(1,10000).__str__() + '.xls'
                    columns = [
                        (u"CEDULA", 3000),
                        (u"APELLIDOS", 5000),
                        (u"NOMBRES", 5000),
                        (u"CARRERA", 5000),
                        (u"CORREO ELECTRONICO", 5000),
                        (u"NÚMERO TELEFÓNICO", 5000),
                        (u"PROMEDIO", 5000),
                        (u"ESTADO", 5000),
                    ]
                    row_num = 4
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    date_format = xlwt.XFStyle()
                    date_format.num_format_str = 'yyyy/mm/dd'
                    row_num = 5
                    for matricula in matriculas:
                        promedio = 0
                        campo1 = str(matricula.inscripcion.persona.identificacion())
                        campo2 = str(matricula.inscripcion.persona.apellido1 + " " + matricula.inscripcion.persona.apellido2)
                        campo3 = str(matricula.inscripcion.persona.nombres)
                        campo4 = str(matricula.inscripcion.carrera.nombre_completo())
                        campo5 = str(matricula.inscripcion.persona.lista_emails())
                        campo6 = str(matricula.inscripcion.persona.telefono)
                        promedio =  MateriaAsignada.objects.filter(status=True, matricula=matricula, materia__esintroductoria=False).aggregate(promedio=Avg('notafinal'))['promedio']
                        estadofinal = "APROBADO"
                        materiasasignadas = MateriaAsignada.objects.values_list('id', flat=True).filter(status=True,
                                                                                                        notafinal__lt=69.5,
                                                                                                        matricula=matricula,
                                                                                                        materia__esintroductoria=False)
                        if materiasasignadas.__len__() > 0:
                            estadofinal = "REPROBADO"
                        ws.write(row_num, 0, campo1, font_style2)
                        ws.write(row_num, 1, campo2, font_style2)
                        ws.write(row_num, 2, campo3, font_style2)
                        ws.write(row_num, 3, campo4, font_style2)
                        ws.write(row_num, 4, campo5, font_style2)
                        ws.write(row_num, 5, campo6, font_style2)
                        ws.write(row_num, 6, promedio, font_style2)
                        ws.write(row_num, 7, estadofinal, font_style2)
                        row_num += 1
                    wb.save(response)
                    return response
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos. %s" % ex})

            elif action == 'listado_alumnos_aprobados_presencial':
                try:
                    __author__ = 'Unemi'
                    matriculas = Matricula.objects.filter(status=True,nivel__periodo=periodo, inscripcion__modalidad__id__in = [1,2], inscripcion__carrera__coordinacion__id=9).distinct().order_by( 'inscripcion__persona__apellido1', 'inscripcion__persona__apellido2','inscripcion__persona__nombres')
                    title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('alumnos_admision_online')
                    ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=listado_alumnos_aprobados_presencial' + random.randint(1,10000).__str__() + '.xls'
                    columns = [
                        (u"CEDULA", 3000),
                        (u"APELLIDOS", 5000),
                        (u"NOMBRES", 5000),
                        (u"CARRERA", 5000),
                        (u"CORREO ELECTRONICO", 5000),
                        (u"NÚMERO TELEFÓNICO", 5000),
                        (u"PROMEDIO", 5000),
                        (u"ESTADO", 5000),
                    ]
                    row_num = 4
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    date_format = xlwt.XFStyle()
                    date_format.num_format_str = 'yyyy/mm/dd'
                    row_num = 5
                    for matricula in matriculas:
                        promedio = 0
                        campo1 = str(matricula.inscripcion.persona.identificacion())
                        campo2 = str(matricula.inscripcion.persona.apellido1 + " " + matricula.inscripcion.persona.apellido2)
                        campo3 = str(matricula.inscripcion.persona.nombres)
                        campo4 = str(matricula.inscripcion.carrera.nombre_completo())
                        campo5 = str(matricula.inscripcion.persona.lista_emails())
                        campo6 = str(matricula.inscripcion.persona.telefono)
                        promedio =  MateriaAsignada.objects.filter(status=True, matricula=matricula, materia__esintroductoria=False).aggregate(promedio=Avg('notafinal'))['promedio']
                        estadofinal = "APROBADO"
                        materiasasignadas = MateriaAsignada.objects.values_list('id', flat=True).filter(status=True,
                                                                                                        notafinal__lt=69.5,
                                                                                                        matricula=matricula,
                                                                                                        materia__esintroductoria=False)
                        if materiasasignadas.__len__() > 0:
                            estadofinal = "REPROBADO"
                        ws.write(row_num, 0, campo1, font_style2)
                        ws.write(row_num, 1, campo2, font_style2)
                        ws.write(row_num, 2, campo3, font_style2)
                        ws.write(row_num, 3, campo4, font_style2)
                        ws.write(row_num, 4, campo5, font_style2)
                        ws.write(row_num, 5, campo6, font_style2)
                        ws.write(row_num, 6, promedio, font_style2)
                        ws.write(row_num, 7, estadofinal, font_style2)
                        row_num += 1
                    wb.save(response)
                    return response
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos. %s" % ex})

            elif action == 'listado_alumnos_online_puros':
                try:
                    __author__ = 'Unemi'
                    matriculas = Matricula.objects.filter(status=True, nivel__periodo=periodo, inscripcion__modalidad__id=3).distinct().order_by( 'inscripcion__persona__apellido1', 'inscripcion__persona__apellido2','inscripcion__persona__nombres')
                    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
                    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                    title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    style1 = easyxf(num_format_str='D-MMM-YY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('alumnos_admision_online')
                    ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=alumnos_admision_online' + random.randint(1,10000).__str__() + '.xls'
                    columns = [
                        (u"Cedula .", 5000),
                        (u"Carrera .", 5000),
                        (u"Nombres .", 5000),
                        (u"Apellidos .", 5000),
                        (u"ppl .", 5000),
                        (u"email", 5000),
                        (u"email institucional", 5000),
                        (u"Celular", 5000),
                        (u"Codigo pais", 5000),
                        (u"Pais residencia", 5000),
                        (u"Codigo provincia", 5000),
                        (u"Provincia residencia", 5000),
                        (u"Codigo cantón", 5000),
                        (u"Cantón residencia", 5000),
                        (u"Parroquia residencia", 5000),
                        (u"Confirmó ficha", 5000),
                        (u"Cantidad evaluaciones", 5000),
                        (u"Discapacidad", 5000),
                        (u"Porcentaje discapacidad", 5000),
                        (u"Carnet", 5000),
                        (u"Deserto", 5000),
                        (u"Asignatura", 5000),
                        (u"Procentaje cumplimiento sobre 70", 5000),
                    ]
                    row_num = 4
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    date_format = xlwt.XFStyle()
                    date_format.num_format_str = 'yyyy/mm/dd'
                    row_num = 5
                    for matricula in matriculas:
                        campo1 = str(matricula.inscripcion.persona.identificacion())
                        campo2 = str(matricula.inscripcion.carrera.nombre_completo())
                        campo3 = str(matricula.inscripcion.persona.apellido1 + " " + matricula.inscripcion.persona.apellido2)
                        campo4 = str(matricula.inscripcion.persona.nombres)
                        campo5 = str(matricula.inscripcion.persona.email)
                        campo6 = str(matricula.inscripcion.persona.emailinst)
                        campo7 = str(matricula.inscripcion.persona.telefono) if matricula.inscripcion.persona.telefono else "S/N"
                        campo8 = str(matricula.inscripcion.persona.pais) if matricula.inscripcion.persona.pais else "S/N"
                        campo9 = str(matricula.inscripcion.persona.provincia) if matricula.inscripcion.persona.provincia else "S/N"
                        campo10 = str(matricula.inscripcion.persona.canton) if matricula.inscripcion.persona.canton else "S/N"
                        campo11 = str(matricula.inscripcion.persona.parroquia) if matricula.inscripcion.persona.parroquia else "S/N"
                        campo12 = str("SI") if matricula.inscripcion.persona.tiene_ficha_confirmada() else "NO"
                        campo14 = str(matricula.inscripcion.persona.mi_perfil().tipodiscapacidad) if matricula.inscripcion.persona.mi_perfil().tipodiscapacidad else "S/N"
                        campo15 = str(matricula.inscripcion.persona.mi_perfil().porcientodiscapacidad) if matricula.inscripcion.persona.mi_perfil().porcientodiscapacidad else "0"
                        campo16 = str( matricula.inscripcion.persona.mi_perfil().carnetdiscapacidad) if matricula.inscripcion.persona.mi_perfil().carnetdiscapacidad else "S/N"
                        campo19 = "SI" if matricula.inscripcion.desertaonline else "NO"
                        campo20 = "SI" if matricula.inscripcion.persona.ppl else "NO"
                        campo21 = str(matricula.inscripcion.persona.pais.id) if matricula.inscripcion.persona.pais else "S/N"
                        campo22 = str(matricula.inscripcion.persona.provincia.id) if matricula.inscripcion.persona.provincia else "S/N"
                        campo23 = str(matricula.inscripcion.persona.canton.id) if matricula.inscripcion.persona.canton else "S/N"
                        campo27 = matricula.cantidad_evaluacion_estudiantes_realizada_periodo(periodo)
                        try:
                            campo28 = matricula.inscripcion.porcentaje_total_estudiante_actividad()
                        except Exception as ex:
                            campo28 = 0
                        ws.write(row_num, 0, campo1, font_style2)
                        ws.write(row_num, 1, campo2, font_style2)
                        ws.write(row_num, 2, campo3, font_style2)
                        ws.write(row_num, 3, campo4, font_style2)
                        ws.write(row_num, 4, campo20, font_style2)
                        ws.write(row_num, 5, campo5, font_style2)
                        ws.write(row_num, 6, campo6, font_style2)
                        ws.write(row_num, 7, campo7, font_style2)
                        ws.write(row_num, 8, campo21, font_style2)
                        ws.write(row_num, 9, campo8, font_style2)
                        ws.write(row_num, 10, campo22, font_style2)
                        ws.write(row_num, 11, campo9, font_style2)
                        ws.write(row_num, 12, campo23, font_style2)
                        ws.write(row_num, 13, campo10, font_style2)
                        ws.write(row_num, 14, campo11, font_style2)
                        ws.write(row_num, 15, campo12, font_style2)
                        ws.write(row_num, 16, campo27, font_style2)
                        ws.write(row_num, 17, campo14, font_style2)
                        ws.write(row_num, 18, campo15, font_style2)
                        ws.write(row_num, 19, campo16, font_style2)
                        ws.write(row_num, 20, campo19, font_style2)
                        ws.write(row_num, 22, campo28, font_style2)
                        mismaterias = matricula.mis_materias_sin_retiro()
                        for materia in mismaterias:
                            campo17 = materia.materia.paralelo
                            ws.write(row_num, 21, campo17, font_style2)
                            break
                        row_num += 1
                    wb.save(response)
                    return response
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos. %s" % ex})

            elif action == 'listado_alumnos_matriculados':
                try:
                    __author__ = 'Unemi'

                    matriculas = Matricula.objects.filter(status=True,nivel__periodo=periodo, inscripcion__modalidad__id=3).distinct().order_by( 'inscripcion__persona__apellido1', 'inscripcion__persona__apellido2','inscripcion__persona__nombres')
                    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
                    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                    title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    style1 = easyxf(num_format_str='D-MMM-YY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('alumnos_admision_online')
                    ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=alumnos_matriculados_online' + random.randint(1,10000).__str__() + '.xls'
                    columns = [
                        (u"Cedula .", 5000),
                        (u"Apellidos .", 5000),
                        (u"Nombres .", 5000),
                        (u"Sexo .", 5000),
                        (u"Usuario .", 5000),
                        (u"email institucional", 5000),
                        (u"email", 5000),
                        (u"Celular sin Codigo", 5000),
                        (u"Celular con Codigo", 5000),
                        (u"Carrera", 5000),
                    ]
                    row_num = 4
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    date_format = xlwt.XFStyle()
                    date_format.num_format_str = 'yyyy/mm/dd'
                    row_num = 5
                    for matricula in matriculas:
                        campo1 = str(matricula.inscripcion.persona.identificacion())
                        campo2 = str(matricula.inscripcion.persona.apellido1 + " " + matricula.inscripcion.persona.apellido2)
                        campo3 = str(matricula.inscripcion.persona.nombres)
                        campo4 = str(matricula.inscripcion.persona.sexo.nombre ) if not matricula.inscripcion.persona.sexo == None else "S/N"
                        campo5 = str(matricula.inscripcion.persona.usuario.username)if not matricula.inscripcion.persona.usuario == None else "S/N"
                        campo6 = str(matricula.inscripcion.persona.emailinst) if matricula.inscripcion.persona.emailinst else "S/N"
                        campo7 = str(matricula.inscripcion.persona.email) if matricula.inscripcion.persona.email else "S/N"
                        campo10 = matricula.inscripcion.carrera.nombre_completo()
                        if matricula.inscripcion.persona.telefono:
                            telefono1=matricula.inscripcion.persona.telefono
                            if len(telefono1)==10 and telefono1[:1]=='0':
                                telefono = telefono1.strip(" ").strip("0")
                                campo8="593"+str(telefono)
                                campo9="+593"+str(telefono)
                            else:
                                campo8 = telefono
                                campo9 = telefono
                        else:
                            campo8 = "N/A"
                            campo9 = "N/A"
                        ws.write(row_num, 0, campo1, font_style2)
                        ws.write(row_num, 1, campo2, font_style2)
                        ws.write(row_num, 2, campo3, font_style2)
                        ws.write(row_num, 3, campo4, font_style2)
                        ws.write(row_num, 4, campo5, font_style2)
                        ws.write(row_num, 5, campo6, font_style2)
                        ws.write(row_num, 6, campo7, font_style2)
                        ws.write(row_num, 7, campo8, font_style2)
                        ws.write(row_num, 8, campo9, font_style2)
                        ws.write(row_num, 9, campo10, font_style2)
                        row_num += 1
                    wb.save(response)
                    return response
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos. %s" % ex})

            elif action == 'listado_docentes_online':
                try:
                    __author__ = 'Unemi'
                    # profesormaterias = ProfesorMateria.objects.filter(tipoprofesor=8,principal=True,activo=True,materia__nivel__periodo=periodo).distinct('profesor__persona__id')
                    profesormaterias = ProfesorMateria.objects.filter(tipoprofesor=8,principal=True,activo=True,materia__nivel__periodo=periodo).distinct('profesor__persona__id')

                    total_enviados = 0
                    total_calificados = 0
                    porcentaje_tareas = 0
                    total_foros_aportados = 0
                    total_foros_calificados = 0
                    total_por_calificar_tareas=0
                    total_por_calificar_foros=0
                    porcentaje_foros = 0
                    hoy = datetime.now().date()
                    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on',num_format_str='#,##0.00')
                    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                    title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    style1 = easyxf(num_format_str='D-MMM-YY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('listado_docente_online')
                    ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=listado_docente_online' + random.randint(1, 10000).__str__() + '.xls'
                    columns = [
                        (u"Cedula .", 2000),
                        (u"Nombres .", 2000),
                        (u"Apellidos .", 2000),
                        (u"Carrera", 2000),
                        (u"Asignatura", 2000),
                        (u"Paralelo", 2000),
                        (u"Total matriculados", 2000),
                        (u"Total evaluaron", 2000),
                        (u"Calificación evaluacion", 2000),
                        (u"Total de Tareas Enviadas", 5000),
                        (u"Total de Tareas Calificadas", 5000),
                        (u"Porcentaje de Tareas", 5000),
                        (u"Total de Foros Participados", 5000),
                        (u"Total de Foros Calificadas", 5000),
                        (u"Porcentaje de Foros", 5000),
                    ]
                    row_num = 4
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    date_format = xlwt.XFStyle()
                    date_format.num_format_str = 'yyyy/mm/dd'
                    row_num = 5
                    for profesormateria in profesormaterias:
                        campo1 = str(profesormateria.profesor.persona.identificacion())
                        campo2 = str(profesormateria.profesor.persona.nombres)
                        campo3 = str(profesormateria.profesor.persona.apellido1 + " " + profesormateria.profesor.persona.apellido2)
                        campo4= str(profesormateria.materia.asignaturamalla.malla.carrera.nombre_completo())
                        campo5= str(profesormateria.materia.asignatura.nombre)
                        campo8= str(profesormateria.materia.paralelomateria.nombre)
                        campo9= profesormateria.materia.cantidad_total_matriculas_materia(profesormateria.profesor,periodo)
                        campo10 = profesormateria.cantidad_estudiantes_encuestados_docencia()
                        campo11 = profesormateria.profesor.promedio_estudiantes_coordinacion_docencia(periodo,coordinacion)
                        resultados_tareas = profesormateria.extraer_tareas()
                        for tarea in resultados_tareas:
                            id_tarea = tarea[0]
                            id_xml_tarea = tarea[1]
                            xmltext = minidom.parseString(id_xml_tarea)
                            itemlist = xmltext.getElementsByTagName("assignment")
                            fecha_fin_sinformato = itemlist[0].getAttribute('duedate')
                            fecha_fin = convertir_fecha(u'%s-%s-%s' % (fecha_fin_sinformato[6:8], fecha_fin_sinformato[4:6], fecha_fin_sinformato[0:4]))
                            restardias = timedelta(days=-1)
                            fecha_fin = fecha_fin + restardias
                            if (fecha_fin <= hoy):
                                campo6,campo7,pendiente = profesormateria.extraer_tareas_enviadas_calificadas(id_tarea)
                                total_enviados += campo6
                                total_calificados += campo7
                        if total_enviados>0:
                            porcentaje_tareas = null_to_decimal((total_calificados*100)/total_enviados,2)

                        # total_foros_aportados = profesormateria.extraer_foros_a_calificar()
                        # total_foros_calificados = profesormateria.extraer_foros_calificados()
                        # if total_foros_aportados >0  and total_foros_calificados >0:
                        #     if total_foros_calificados > total_foros_aportados:
                        #         porcentaje_foros = 100
                        #     else:
                        #         porcentaje_foros = null_to_decimal((total_foros_calificados * 100) / total_foros_aportados, 2)

                        # for foro in resultados_foros:
                        #     total_foros_aportados += foro[7]
                            # if not foro[6] == None:
                            #     fecha_finalizacion = convertir_fecha(foro[6].strftime('%d-%m-%Y'))
                            #     restardias = timedelta(days=-1)
                            #     fecha_fin = fecha_finalizacion + restardias
                            #     if fecha_fin < hoy:
                            #         resultado_foros_calificados = profesormateria.extraer_foros_calificados()
                                    # for calificados in resultado_foros_calificados:
                                    #     if not calificados[0]== None:
                                    #         if calificados[0] >= 0:
                                    #             if calificados[1] == foro[1] or foro[2] in calificados[1]:
                                    #                 total_foros_calificados += 1
                                    #             else:
                                    #                 titulo_foro =  foro[3].replace(u'.', '')
                                    #                 if calificados[1] == foro[4] or titulo_foro in calificados[1]:
                                    #                     total_foros_calificados += 1

                        ws.write(row_num, 0, campo1, font_style2)
                        ws.write(row_num, 1, campo2, font_style2)
                        ws.write(row_num, 2, campo3, font_style2)
                        ws.write(row_num, 3, campo4, font_style2)
                        ws.write(row_num, 4, campo5, font_style2)
                        ws.write(row_num, 5, campo8, font_style2)
                        ws.write(row_num, 6, campo9, font_style2)
                        ws.write(row_num, 7, campo10, font_style2)
                        ws.write(row_num, 8, campo11, font_style2)
                        ws.write(row_num, 9, total_enviados, font_style2)
                        ws.write(row_num, 10, total_calificados, font_style2)
                        ws.write(row_num, 11, porcentaje_tareas, font_style2)
                        ws.write(row_num, 12, total_foros_aportados, font_style2)
                        ws.write(row_num, 13, total_foros_calificados, font_style2)
                        ws.write(row_num, 14, profesormateria.porcentaje_cumplimiento_foros(), font_style2)
                        row_num += 1
                        total_enviados = 0
                        total_calificados = 0
                        total_foros_calificados = 0
                        total_foros_aportados = 0
                        porcentaje_foros=0
                        porcentaje_tareas=0
                    wb.save(response)
                    return response
                except Exception as ex:
                    pass

            elif action == 'informe_actividades_docente_tutor':
                try:
                    __author__ = 'Unemi'
                    profesormaterias = ProfesorMateria.objects.filter(tipoprofesor=8,principal=True,activo=True,materia__nivel__periodo=periodo).distinct('profesor__persona__id')
                    nombre_tarea=''
                    nombre_foro=''
                    fecha_inicio=''
                    fecha_fin_tarea=''
                    hoy = datetime.now().date()
                    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on',num_format_str='#,##0.00')
                    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                    title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    style1 = easyxf(num_format_str='D-MMM-YY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('listado_docente_online')
                    ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=informe_tutores' + random.randint(1, 10000).__str__() + '.xls'
                    columns = [
                        (u"Cedula .", 3000),
                        (u"Nombres .", 3000),
                        (u"Apellidos .", 3000),
                        (u"Carrera", 3000),
                        (u"Asignatura", 3000),
                        (u"Paralelo", 2000),
                        (u"Actividad", 2000),
                        (u"Nombre de Actividad", 2000),
                        (u"Fecha de Inicio", 2000),
                        (u"Fecha Fin", 2000),
                        (u"# Matriculados", 2000),
                        (u"Total de Actividad Enviadas", 5000),
                        (u"Total de Actividad Calificadas", 5000),
                        (u"Total de Actividad Sin Calificar", 5000),
                    ]
                    row_num = 4
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    date_format = xlwt.XFStyle()
                    date_format.num_format_str = 'yyyy/mm/dd'
                    row_num = 5
                    for profesormateria in profesormaterias:
                        campo1 = str(profesormateria.profesor.persona.identificacion())
                        campo2 = str(profesormateria.profesor.persona.nombres)
                        campo3 = str(profesormateria.profesor.persona.apellido1 + " " + profesormateria.profesor.persona.apellido2)
                        campo4= str(profesormateria.materia.asignaturamalla.malla.carrera.nombre_completo())
                        campo5= str(profesormateria.materia.asignatura.nombre)
                        campo6= str(profesormateria.materia.paralelomateria.nombre)
                        campo7= profesormateria.materia.cantidad_total_matriculas_materia(profesormateria.profesor,periodo)
                        resultados_tareas = profesormateria.extraer_tareas()
                        for tarea in resultados_tareas:
                            id_tarea = tarea[0]
                            id_xml_tarea = tarea[1]
                            xmltext = minidom.parseString(id_xml_tarea)
                            itemlist = xmltext.getElementsByTagName("assignment")
                            fecha_fin_sinformato = itemlist[0].getAttribute('duedate')
                            fecha_fin = convertir_fecha(u'%s-%s-%s' % (fecha_fin_sinformato[6:8], fecha_fin_sinformato[4:6], fecha_fin_sinformato[0:4]))
                            fecha_fin_tarea=fecha_fin
                            fecha_inicio_sinformato = itemlist[0].getAttribute('opendate')
                            fecha_inicio = convertir_fecha(u'%s-%s-%s' % (fecha_inicio_sinformato[6:8], fecha_inicio_sinformato[4:6], fecha_inicio_sinformato[0:4]))
                            restardias = timedelta(days=-1)
                            fecha_fin = fecha_fin + restardias
                            if (fecha_fin <= hoy):
                                enviadas,calificadas,pendientes= profesormateria.extraer_tareas_enviadas_calificadas(id_tarea)
                                if (pendientes > 0):
                                    nombre_tarea = (itemlist[0].getAttribute('title')).upper()
                                    ws.write(row_num, 0, campo1, font_style2)
                                    ws.write(row_num, 1, campo2, font_style2)
                                    ws.write(row_num, 2, campo3, font_style2)
                                    ws.write(row_num, 3, campo4, font_style2)
                                    ws.write(row_num, 4, campo5, font_style2)
                                    ws.write(row_num, 5, campo6, font_style2)
                                    ws.write(row_num, 6, 'TAREAS', font_style2)
                                    ws.write(row_num, 7, nombre_tarea, font_style2)
                                    ws.write(row_num, 8, str(fecha_inicio), font_style2)
                                    ws.write(row_num, 9, str(fecha_fin_tarea), font_style2)
                                    ws.write(row_num, 10, campo7, font_style2)
                                    ws.write(row_num, 11, enviadas, font_style2)
                                    ws.write(row_num, 12, calificadas, font_style2)
                                    ws.write(row_num, 13, pendientes, font_style2)
                                    row_num += 1

                        resultados_foros = profesormateria.extraer_foros()
                        for foro in resultados_foros:
                            foros_calificados = 0
                            foros_por_calificar = 0
                            titulo_topico = foro[1]
                            asigname_topico = foro[2]
                            titulo_foro = foro[3]
                            asigname_foro = foro[4]
                            default_name = 'SI' if not foro[2] == None else 'NO'
                            if foro[6]:
                                fecha_finalizacion = convertir_fecha(foro[6].strftime('%d-%m-%Y'))
                                fecha_apertura = convertir_fecha(foro[5].strftime('%d-%m-%Y'))
                                if (fecha_finalizacion <= hoy):
                                    if default_name == 'SI':
                                        resultado_foros_calificados = profesormateria.extraer_foros_calificados_docente()
                                        for calificados in resultado_foros_calificados:
                                            if not calificados[0] == None:
                                                if calificados[0] >= 0 and (calificados[1] == asigname_topico or titulo_topico in calificados[1]):
                                                    foros_calificados += 1
                                                else:
                                                    titulo_foro = titulo_foro.replace(u'.', '')
                                                    if calificados[0] >= 0 and ((calificados[1] == asigname_foro or titulo_foro in calificados[1])):
                                                        foros_calificados += 1

                                        foros_por_calificar = foro[7] - foros_calificados if not foros_calificados > foro[7] else 0
                                        if foros_por_calificar > 0:
                                            nombre_foro=titulo_topico
                                            ws.write(row_num, 0, campo1, font_style2)
                                            ws.write(row_num, 1, campo2, font_style2)
                                            ws.write(row_num, 2, campo3, font_style2)
                                            ws.write(row_num, 3, campo4, font_style2)
                                            ws.write(row_num, 4, campo5, font_style2)
                                            ws.write(row_num, 5, campo6, font_style2)
                                            ws.write(row_num, 6, 'FOROS', font_style2)
                                            ws.write(row_num, 7, nombre_foro, font_style2)
                                            ws.write(row_num, 8, str(fecha_apertura), font_style2)
                                            ws.write(row_num, 9, str(fecha_finalizacion), font_style2)
                                            ws.write(row_num, 10, campo7, font_style2)
                                            ws.write(row_num, 11, foro[7], font_style2)
                                            ws.write(row_num, 12, foros_calificados, font_style2)
                                            ws.write(row_num, 13, foros_por_calificar, font_style2)
                                            row_num += 1
                    wb.save(response)
                    return response
                except Exception as ex:
                    pass


            elif action == 'notas_migradas':
                try:
                    data['title'] = u'Actividades migradas'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['idinscripcion'])
                    data['materia'] = materia = Materia.objects.get(id=request.GET['idcurso'])
                    return render(request, "inscripciones_admision/notas_migradas.html", data)
                except Exception as ex:
                    transaction.set_rollback(True)
                    pass

            # elif action == 'delactividadsakai':
            #     try:
            #         data['title'] = u'Elminación de actividad'
            #         data['actividad'] = ActividadesSakaiAlumno.objects.get(id=request.GET['id'])
            #         return render(request, "inscripciones_admision/delactividadsakai.html", data)
            #     except Exception as ex:
            #         pass

            elif action == 'listado_matricular':
                try:
                    __author__ = 'Unemi'
                    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
                    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                    title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    style1 = easyxf(num_format_str='D-MMM-YY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('alumnos_admision_online')
                    ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=alumnos_admision_online' + random.randint(1,10000).__str__() + '.xls'
                    columns = [
                        (u"username", 5000),
                        (u"password", 5000),
                        (u"firstname", 5000),
                        (u"lastname", 5000),
                        (u"email", 5000),
                        (u"country", 5000),
                        (u"city", 5000),
                        (u"institucion", 5000),
                        (u"lang", 5000),
                        (u"course1", 5000),
                        (u"course2", 5000),
                        (u"course3", 5000),
                        (u"course4", 5000),
                        (u"course5", 5000),
                    ]
                    row_num = 4
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    date_format = xlwt.XFStyle()
                    date_format.num_format_str = 'yyyy/mm/dd'
                    row_num = 4
                    sql = "select a.cedula as username, b.cedula as password, max(b.apel) as firstname, max(b.nom) as lastname, max(b.email_inst) as email, max(a.sede) as country, max(a.canton) as city, max(a.carrera) as institution, 'es' as lang, max(a.shortname) as course1 from horario as a inner join ficha_estudiante as b on b.cedula=a.cedula group by a.cedula,a.shortname order by a.cedula;"
                    resultados = querymysqlconsulta(sql, True)
                    aux =[]
                    if resultados:
                        columna = 8
                        i=0
                        for matricula in resultados:
                            campo1 = str(matricula[0].strip())
                            if not campo1 in aux:
                                aux.append(campo1)
                                columna = 8
                                row_num += 1
                                campo2 = str(matricula[1].strip())
                                campo3 = str(matricula[2])
                                campo4 = str(matricula[3])
                                campo5 = str(matricula[4])
                                campo6 = str(matricula[5])
                                campo7 = str(matricula[6])
                                campo8 = str(matricula[7])
                                campo9 = str(matricula[8])
                                ws.write(row_num, 0, campo1, font_style2)
                                ws.write(row_num, 1, campo2, font_style2)
                                ws.write(row_num, 2, campo3, font_style2)
                                ws.write(row_num, 3, campo4, font_style2)
                                ws.write(row_num, 4, campo5, font_style2)
                                ws.write(row_num, 5, campo6, font_style2)
                                ws.write(row_num, 6, campo7, font_style2)
                                ws.write(row_num, 7, campo8, font_style2)
                                ws.write(row_num, 8, campo9, font_style2)
                            campo10 = str(matricula[9])
                            columna += 1
                            ws.write(row_num, columna, campo10, font_style2)
                    wb.save(response)
                    return response
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos. %s" % ex})

            elif action == 'migrar_examen_moodle':
                try:
                    data['title'] = u'Importar notas examen'
                    data['form'] = ImportarArchivoXLSForm()
                    return render(request, "inscripciones_admision/migrar_examen_moodle.html", data)
                except Exception as ex:
                    pass

            elif action == 'addactividadsakai':
                try:
                    data['title'] = u'Adicionar actividad'
                    data['inscripcion'] = Inscripcion.objects.get(pk=request.GET['id'])
                    data['materia'] = Materia.objects.get(pk=request.GET['idmat'])
                    form = ActividadSakaiForm()
                    data['form'] = form
                    return render(request, "inscripciones_admision/addactividadsakai.html", data)
                except Exception as ex:
                    pass

            elif action == 'importar_listado_alumnos':
                try:
                    data['title'] = u'Importar Alumnos'
                    data['form'] = ImportarListadoAlumnoForm()
                    return render(request, "inscripciones_admision/importar_listado_alumnos.html", data)
                except Exception as ex:
                    pass

            elif action == 'estudiantes_sin_acceso_plataforma':
                try:
                    data['title'] = u'Alumnos sin acceso a la pltaforma'
                    data['form'] = ImportarListadoAlumnoForm()
                    return render(request, "inscripciones_admision/estudiantes_sin_acceso_plataforma.html", data)
                except Exception as ex:
                    pass

            elif action == 'descargar_formato':
                try:
                    __author__ = 'Unemi'
                    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
                    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                    title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    style1 = easyxf(num_format_str='D-MMM-YY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('exp_xls_post_part')
                    # ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    response = HttpResponse(content_type="application/ms-excel")
                    response[
                        'Content-Disposition'] = 'attachment; filename=formato_matricula' + random.randint(1, 10000).__str__() + '.xls'
                    columns = [
                        (u"CEDULA", 3500),
                        (u"NOMBRES", 10000),
                        (u"APELLIDO 1 ", 10000),
                        (u"APELLIDO 2 ", 10000),
                        (u"EMAIL", 8000),
                        (u"ID CARRERA", 3000),
                        (u"PARALELO", 3000),
                        (u"CARRERA", 8000),
                        (u"ID PAIS", 3000),
                        (u"PAIS", 4000),
                        (u"ID PROVINCIA", 3000),
                        (u"PROVINCIA", 4000),
                        (u"ID CANTON", 3000),
                        (u"CANTON", 4000),
                        (u"CELULAR", 3000),
                        (u"TELEFONO", 3000),
                        (u"FECHA NACIMIENTO", 4000),
                    ]
                    row_num = 0
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    date_format = xlwt.XFStyle()
                    date_format.num_format_str = 'yyyy/mm/dd'
                    # listadoinscripcion = VirtualSoporteUsuarioInscripcion.objects.filter(soporteusuario__persona=persona, activo=True,matricula__nivel__periodo=periodo).order_by('matricula__inscripcion__persona__apellido1')
                    # row_num = 4
                    # for listado in listadoinscripcion:
                    #     i = 0
                    #     campo1 = listado.matricula.inscripcion.persona.cedula
                    #     campo2 = listado.matricula.inscripcion.persona
                    #     campo3 = listado.matricula.inscripcion.persona.email
                    #     campo4 = listado.matricula.inscripcion.persona.telefono
                    #     campo5 = listado.virtualsoporteusuarioincidentes_set.filter(status=True).count()
                    #     campo6 = ''
                    #     campo7 = ''
                    #     if listado.matricula.inscripcion.persona.ppl:
                    #         campo6 = 'SI'
                    #     if listado.matricula.inscripcion.persona.pais:
                    #         campo7 = listado.matricula.inscripcion.persona.pais.nombre
                    #
                    # ws.write(row_num, 0, campo1, font_style2)
                    # ws.write(row_num, 1, campo2.__str__(), font_style2)
                    # ws.write(row_num, 2, campo3, font_style2)
                    # ws.write(row_num, 3, campo4, font_style2)
                    # ws.write(row_num, 4, campo5, font_style2)
                    # ws.write(row_num, 5, campo6, font_style2)
                    # ws.write(row_num, 6, campo7, font_style2)
                    # row_num += 1
                    wb.save(response)
                    return response
                except Exception as ex:
                    pass


            elif action == 'importar_listado_soporte_alumno':
                try:
                    data['title'] = u'Asignar Estudiantes a Soportes'
                    data['form'] = ImportarListadoAlumnoSoporteForm()
                    return render(request, "inscripciones_admision/importar_listado_soporte_alumno.html", data)
                except Exception as ex:
                    pass

            elif action == 'importar_listado_soporte_admin':
                try:
                    data['title'] = u'Asignar Estudiantes a Soportes'
                    data['form'] = ImportarListadoAlumnoSoporteForm()
                    return render(request, "inscripciones_admision/soporte_usuario_admin.html", data)
                except Exception as ex:
                    pass


            elif action == 'soportes':
                try:
                    data['title'] = u'Listado Soportes'
                    if 'idt' in request.GET:
                        data['tutor'] = tutor = Profesor.objects.get(pk=request.GET['idt'])
                    else:
                        data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    data['soportes'] = VirtualSoporteUsuario.objects.filter(status=True)
                    return render(request, "inscripciones_admision/soportes.html", data)
                except Exception as ex:
                    pass

            elif action == 'soportes_activos':
                try:
                    data['title'] = u'Listado de Soportes'
                    # data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    data['soportes'] = VirtualSoporteUsuario.objects.filter(activo=True)
                    return render(request, "inscripciones_admision/soporte_activo.html", data)
                except Exception as ex:
                    pass

            elif action == 'ver_asigandos':
                try:
                    data['title'] = u'Estudiantes Asignados'
                    # data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    data['soporte'] = soporte = VirtualSoporteUsuario.objects.get(pk=request.GET['id'])
                    data['asignados'] = soporte.asignados(periodo)
                    return render(request, "inscripciones_admision/asignar_usuarios.html", data)
                except Exception as ex:
                    pass

            elif action == 'ver_materias_asignadas':
                try:
                    data['title'] = u'Materias Asignadas'
                    data['soporte'] = soporte = VirtualSoporteUsuario.objects.get(pk=request.GET['id'])
                    data['asignaturas'] = Asignatura.objects.filter(materia__nivel__periodo=periodo,materia__nivel__modalidad=3).distinct()
                    return render(request, "inscripciones_admision/ver_asignaturas_gestion.html", data)
                except Exception as ex:
                    pass

            elif action == 'listado_tutores':
                try:
                    data['title'] = u'Listado de Tutores online'
                    search = None
                    # listado_tutores = ProfesorMateria.objects.filter(materia__asignatura__id__in=Materia.objects.values_list('asignatura__id').filter(nivel__id=414).distinct(), tipoprofesor=8, principal=True,materia__nivel__periodo=periodo).distinct('profesor__id')
                    listado_tutores = ProfesorMateria.objects.filter(tipoprofesor=8, principal=True,materia__nivel__periodo=periodo).distinct('profesor__id')
                    if 's' in request.GET:
                        search = request.GET['s'].strip()
                        ss = search.split(' ')
                        if len(ss) == 1:
                            listado_tutores = listado_tutores.filter(Q(profesor__persona__nombres__icontains=search) |
                                                                                     Q(profesor__persona__apellido1__icontains=search) |
                                                                                     Q(profesor__persona__apellido2__icontains=search) |
                                                                                     Q(profesor__persona__cedula__icontains=search) |
                                                                                     Q(profesor__persona__pasaporte__icontains=search) )
                        else:
                            listado_tutores = listado_tutores.filter(Q(profesor__persona__apellido1__icontains=ss[0]) &
                                                                                     Q(profesor__persona__apellido2__icontains=ss[1]))
                    paging = MiPaginador(listado_tutores, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['page'] = page
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['listado_tutores'] = page.object_list
                    data['search'] = search if search else ""
                    return render(request, "inscripciones_admision/tutores_activos.html", data)
                except Exception as ex:
                    pass

            elif action == 'ver_soportes_asigandos':
                try:
                    data['title'] = u'Soportes Asignados'
                    data['soporte'] = soporte = VirtualSoporteUsuario.objects.get(pk=request.GET['id'])
                    data['asignados'] = soporte.asignados(periodo)
                    return render(request, "inscripciones_admision/asignar_usuarios.html", data)
                except Exception as ex:
                    pass

            elif action == 'ppl':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_inscripciones')
                    data['title'] = u'Historial PPL'
                    search = None
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=int(encrypt(request.GET['id'])))
                    historialppl = HistorialPersonaPPL.objects.filter(persona=inscripcion.persona, status=True).order_by('fechaingreso')
                    paging = MiPaginador(historialppl, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['historialppl'] = page.object_list
                    return render(request, "inscripciones_admision/historialppl.html", data)
                except Exception as ex:
                    return HttpResponseRedirect(f'{request.path}?info={ex.__str__()}')

            elif action == 'addppl':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_inscripciones')
                    data['title'] = u'Adicionar registro de PPL al Aspirante'
                    data['id'] = int(encrypt(request.GET['idi']))
                    data['form'] = PersonaPPLForm()
                    return render(request, "inscripciones_admision/addppl.html", data)
                except Exception as ex:
                    pass

            elif action == 'editppl':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_inscripciones')
                    data['title'] = u'Editar registro de PPL al Aspirante'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=int(encrypt(request.GET['idi'])))
                    data['hppl'] = hppl = HistorialPersonaPPL.objects.get(pk=int(encrypt(request.GET['id'])))
                    form = PersonaPPLForm(initial={'fechaingresoppl': hppl.fechaingreso,
                                                   'fechasalidappl': hppl.fechasalida,
                                                   'observacionppl': hppl.observacion,
                                                   'archivoppl': None,
                                                   'centrorehabilitacion': hppl.centrorehabilitacion,
                                                   'lidereducativo': hppl.lidereducativo,
                                                   'correolidereducativo': hppl.correolidereducativo,
                                                   'telefonolidereducativo': hppl.telefonolidereducativo,
                                                   })

                    data['form'] = form
                    return render(request, "inscripciones_admision/editppl.html", data)
                except Exception as ex:
                    pass

            elif action == 'deleteppl':
                try:
                    puede_realizar_accion(request, 'sga.puede_eliminar_ppl')
                    data['title'] = u'Eliminar registro de PPL al Aspirante'
                    data['inscripcion'] = Inscripcion.objects.get(pk=request.GET['idi'])
                    data['hppl'] = HistorialPersonaPPL.objects.get(pk=request.GET['id'])
                    return render(request, "inscripciones_admision/deleteppl.html", data)
                except Exception as ex:
                    pass

            elif action == 'reporteaula':
                try:
                    periodo = request.GET['periodo']
                    # __author__ = 'Unemi'
                    # style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                    # style_nb = easyxf('font: name Times New Roman, color-index blue, bold on',
                    #                   num_format_str='#,##0.00')
                    # style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                    # title = easyxf(
                    #     'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    # style1 = easyxf(num_format_str='D-MMM-YY')
                    # font_style = XFStyle()
                    # font_style.font.bold = True
                    # font_style2 = XFStyle()
                    # font_style2.font.bold = False
                    # wb = Workbook(encoding='utf-8')
                    # ws = wb.add_sheet('exp_xls_post_part')
                    # ws.write_merge(0, 0, 0, 12, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    # response = HttpResponse(content_type="application/ms-excel")
                    # response[
                    #     'Content-Disposition'] = 'attachment; filename=exp_xls_post_part_' + random.randint(1, 10000).__str__() + '.xls'

                    output = io.BytesIO()
                    workbook = xlsxwriter.Workbook(output)
                    ws = workbook.add_worksheet('Listado')

                    merge_format = workbook.add_format({
                        'bold': 1,
                        'border': 1,
                        'align': 'center',
                        'valign': 'vcenter',
                        'bg_color': 'silver',
                        'text_wrap': 1})

                    font_style2 = workbook.add_format({
                        'border': 1
                    })

                    ws.merge_range(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO',
                                   workbook.add_format({'align': 'center',
                                                        'valign': 'vcenter',
                                                        'bold': 1,
                                                        'font_size': 16}))

                    wb = Workbook(encoding='utf-8')
                    response = HttpResponse(content_type="application/ms-excel")
                    response[
                        'Content-Disposition'] = 'attachment; filename=estudiantessingrupospracticas_' + random.randint(
                        1, 10000).__str__() + '.xls'

                    columns = [
                        (u"CEDULA", 20),
                        (u"DOCENTE", 20),
                        (u"FACULTAD", 20),
                        (u"CARRERA", 20),
                        (u"PARALELO", 20),
                        (u"ASIGNATURA", 20),
                        (u"INICIO HORARIO", 20),
                        (u"FIN HORARIO", 20),
                        (u"HORARIOS", 20),

                        (u"BLOQUE-AULA", 20),

                        # 1(u"CAPACIDAD AULA", 6000),
                        # 2(u"CUPO MATRICULA", 6000),
                        # 3(u"MATRICULADOS", 6000),
                        # 4(u"INSCRITOS", 6000),
                        # 5(u"DÍAS", 6000),
                        #
                        # 10(u"JORNADA", 6000),
                        # 11(u"NIVEL", 6000),
                        #
                        # 15(u"TIPO PROFESOR", 6000),
                        # 16(u"CATEGORIA", 6000),
                        # 17(u"DEDICACIÓN", 6000),
                        # 18(u"PRINCIPAL", 6000),
                        # 20(u"TIPO MATERIA", 6000),
                        # 21(u"ID HORARIO", 4000),
                        # 22(u"INICIO MATERIA", 4000),
                        # 23(u"FIN MATERIA", 4000),

                    ]
                    row_num = 3
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], merge_format)
                        ws.set_column(col_num, col_num, columns[col_num][1])

                    cursor = connections['sga_select'].cursor()
                    # lista_json = []
                    # data = {}
                    sql = f"""
                        SELECT al.nombre AS Aula, al.capacidad AS capacidad_aula, mat.cupo AS cupo_matriculas, 
                         (
                        SELECT COUNT(*)
                        FROM sga_materiaasignada mas1, sga_matricula mat1, sga_nivel ni1
                        WHERE mat1.estado_matricula in (2,3) AND mas1.matricula_id=mat1.id AND mat1.nivel_id=ni1.id AND ni1.periodo_id=ni.periodo_id AND mas1.materia_id=mat.id) AS Matriculados,
                         (
                            SELECT COUNT(*)
                            FROM sga_materiaasignada mas1, sga_matricula mat1, sga_nivel ni1
                            WHERE mat1.estado_matricula=1 AND mas1.matricula_id=mat1.id AND mat1.nivel_id=ni1.id 
                            AND ni1.periodo_id=ni.periodo_id AND mas1.materia_id=mat.id) AS inscritos, 
                         (CASE cl.dia WHEN 1 THEN 'LUNES' WHEN 2 THEN 'MARTES' WHEN 3 THEN 'MIERCOLES' WHEN 4 THEN 'JUEVES' WHEN 5 THEN 'VIERNES' WHEN 6 THEN 'SABADO' WHEN 7 THEN 'DOMINGO' END) AS dia, 
                         cl.inicio, cl.fin, (tu.comienza|| '  ' || tu.termina) AS Horario, asi.nombre AS Asignatura, 

                         ni.paralelo AS jornada, 

                         nmall.nombre AS nivel, 

                         ca.nombre AS Carrera, mat.paralelo AS paralelo, per.apellido1 ||' '|| per.apellido2 ||' '|| per.nombres AS docente,
                         tipop.nombre AS tipo_profesor, cat.nombre AS categorizacion,ded.nombre AS dedicacion,
                         (CASE pm.principal WHEN TRUE THEN 'SI' ELSE 'NO' END) AS principal, cor.nombre AS facultad, (CASE asimall.practicas WHEN TRUE THEN 'SI' ELSE 'NO' END) AS tipomateria, cl.id AS Id, mat.inicio AS Inicio_materia, mat.fin AS Fin_materia, mat.id AS id_materia, per.cedula AS ced
                        FROM sga_clase cl, sga_aula al, sga_materia mat, sga_nivel ni, 
                        sga_asignatura asi,sga_asignaturamalla asimall, sga_nivelmalla nmall, 
                        sga_malla mall, sga_carrera ca, sga_turno tu, sga_profesormateria pm, sga_tipoprofesor tipop, sga_categorizaciondocente cat, sga_tiempodedicaciondocente ded,
                         sga_profesor pr, sga_persona per, sga_coordinacion_carrera cc, sga_coordinacion cor
                        WHERE cl.aula_id=al.id AND cl.materia_id=mat.id AND mat.nivel_id=ni.id AND mat.asignatura_id=asi.id 
                        AND mat.asignaturamalla_id = asimall.id AND pr.dedicacion_id=ded.id AND 
                         pm.materia_id=mat.id AND pr.id=pm.profesor_id AND per.id=pr.persona_id AND asimall.nivelmalla_id = nmall.id 
                         AND asimall.malla_id = mall.id AND tu.id=cl.turno_id AND mall.carrera_id=ca.id AND cc.carrera_id=ca.id 
                         AND cl.activo= TRUE AND cor.id=cc.coordinacion_id AND tipop.id=pm.tipoprofesor_id 
                         AND pr.categoria_id=cat.id AND ni.periodo_id={periodo} 
                    """
                    cursor.execute(sql)
                    results = cursor.fetchall()
                    row_num = 4
                    for r in results:
                        i = 0
                        campo1 = r[0]
                        campo2 = r[1]
                        campo3 = r[2]
                        campo4 = r[3]
                        campo5 = r[4]
                        campo6 = r[5]
                        campo7 = r[6]
                        campo8 = r[7]
                        campo9 = r[8]
                        campo10 = r[9]
                        campo11 = r[10]
                        campo12 = r[11]
                        campo13 = r[12]
                        campo14 = r[13]
                        campo15 = r[14]
                        campo16 = r[15]
                        campo17 = r[16]
                        campo18 = r[17]
                        campo19 = r[18]
                        campo20 = r[19]
                        campo21 = r[20]
                        campo22 = r[21]
                        campo23 = r[22]
                        campo24 = r[23]
                        campo25 = r[25]

                        ws.write(row_num, 0, u"%s" % campo25, font_style2)
                        ws.write(row_num, 1, u"%s" % campo15, font_style2)
                        ws.write(row_num, 2, u"%s" % campo20, font_style2)
                        ws.write(row_num, 3, u"%s" % campo13, font_style2)
                        ws.write(row_num, 4, u"%s" % campo14, font_style2)
                        ws.write(row_num, 5, u"%s" % campo10, font_style2)
                        ws.write(row_num, 6, u"%s" % campo7, font_style2)
                        ws.write(row_num, 7, u"%s" % campo8, font_style2)
                        ws.write(row_num, 8, u"%s" % campo9, font_style2)
                        ws.write(row_num, 9, u"%s" % campo1, font_style2)

                        row_num += 1

                    workbook.close()
                    output.seek(0)
                    # Set up the Http response.
                    filename = 'HORARIOS Y AULAS' + random.randint(1, 10000).__str__() + '.xlsx'
                    response = HttpResponse(output,
                                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = 'attachment; filename=%s' % filename
                    return response
                except Exception as ex:
                    pass

            elif action == 'totalmatriculadossinmodulos':
                try:
                    cursor = connections['sga_select'].cursor()
                    response = HttpResponse(content_type="application/ms-excel")
                    response[
                        'Content-Disposition'] = 'attachment; filename=MATRICULADOS DE ' + periodo.nombre.__str__() + '.xls'
                    periodoid = request.GET['periodo']
                    wb = xlwt.Workbook()
                    ws = wb.add_sheet('Sheetname')
                    fmt = xlwt.easyxf
                    encabezado = fmt(
                        'font: height 150, bold on; border: left thin, right thin, top thin, bottom thin; align: wrap on, vert centre, horiz center;')

                    estilo = xlwt.easyxf(
                        'font: height 350, name Arial, colour_index black, bold on, italic on; align: wrap on, vert centre, horiz center;')
                    ws.write_merge(0, 0, 0, 9, 'UNIVERSIDAD ESTATAL DE MILAGRO', estilo)
                    ws.col(0).width = 1000
                    ws.col(1).width = 6000
                    ws.col(2).width = 3000
                    ws.col(3).width = 3000
                    ws.col(4).width = 6000
                    ws.col(5).width = 6000
                    ws.col(6).width = 6000
                    ws.col(7).width = 6000
                    ws.col(8).width = 4000
                    ws.col(9).width = 6000
                    ws.col(10).width = 6000
                    ws.col(11).width = 6000
                    ws.col(14).width = 4000
                    ws.col(15).width = 4000
                    ws.col(16).width = 4000
                    ws.write(4, 0, 'N.', encabezado)
                    # ws.write(4, 1, 'PERIODO')
                    # ws.write(4, 2, 'NIVEL_CRE')
                    # ws.write(4, 3, 'NIVEL_MAT')
                    # ws.write(4, 4, 'SECCION')
                    ws.write(4, 1, 'CEDULA', encabezado)
                    ws.write(4, 2, 'APELLIDOS', encabezado)
                    ws.write(4, 3, 'NOMBRES', encabezado)
                    ws.write(4, 4, 'CARRERA', encabezado)
                    ws.write(4, 5, 'MODALIDAD', encabezado)
                    ws.write(4, 6, 'NÚMERO DE MATRICULA', encabezado)
                    ws.write(4, 7, 'SEXO', encabezado)
                    ws.write(4, 8, 'FECHANACIMIENTO', encabezado)
                    ws.write(4, 9, 'EMAIL', encabezado)
                    # ws.write(4, 11, 'EMAILINST', encabezado)
                    # ws.write(4, 12, 'COORDINACION', encabezado)
                    # ws.write(4, 14, 'COD. SENESCYT')
                    ws.write(4, 10, 'TELEFONO', encabezado)
                    ws.write(4, 11, 'TIENE DISCAPACIDAD', encabezado)
                    ws.write(4, 12, 'DISCAPACIDAD', encabezado)
                    # ws.write(4, 16, 'USUARIO')
                    # ws.write(4, 16, 'INSCRIPCION')
                    ws.write(4, 13, 'LGTBI', encabezado)
                    ws.write(4, 14, 'ETNIA', encabezado)
                    ws.write(4, 15, 'PPL', encabezado)

                    ws.write(4, 16, 'NACIONALIDAD', encabezado)
                    ws.write(4, 17, 'PAIS DE RESIDENCIA', encabezado)
                    ws.write(4, 18, 'PROVINCIA DE RESIDENCIA', encabezado)
                    ws.write(4, 19, 'CANTON DE RESIDENCIA', encabezado)
                    ws.write(4, 20, 'DIRECCION DOMICILIARIA', encabezado)
                    ws.write(4, 21, 'ESTADO SOCIO ECONOMICO', encabezado)
                    # ws.write(4, 26, 'REALIZADO')
                    # ws.write(4, 26, 'HORAS PRACTICAS')
                    # ws.write(4, 27, 'HORAS VINCULACION')
                    # ws.write(4, 27, 'FECHA INICIO PRIMER NIVEL')
                    # ws.write(4, 28, 'FECHA CONVALIDACION')
                    # ws.write(4, 28, 'PROMEDIO ASISTENCIA')
                    # ws.write(4, 29, 'PROMEDIO NOTAS')
                    # ws.write(4, 30, 'ULTIMO MODULO')
                    # ws.write(4, 31, 'MATRICULADO MODULO')

                    # ws.write(4, 30, 'ITINETARIOS CUMPLIDOS')
                    # ws.write(4, 31, 'ID MATRICULA')
                    # ws.write(4, 32, 'TIPO MATRICULA')
                    # ws.write(4, 33, 'TIPO ESTUDIANTE')
                    ws.write(4, 22, 'CONTACTO EMERGENCIA', encabezado)
                    ws.write(4, 23, 'COLEGIO', encabezado)
                    # ws.write(4, 38, 'ACEPTA MATRÍCULA')
                    ws.write(4, 24, 'GRATUIDAD', encabezado)

                    a = 4
                    date_format = xlwt.XFStyle()
                    date_format.num_format_str = 'yyyy/mm/dd'
                    listaestudiante = """
                                    select peri.nombre as PERIODO, nimalla.nombre as NIVEL_CRE,sesion.nombre as SECCION, 
                                    perso.cedula, perso.apellido1 || ' ' || perso.apellido2 as apellidos, 
                                    perso.nombres as nompersona, se.nombre as sexo, 
                                    perso.nacimiento as fechanacimiento,perso.email,perso.emailinst, coor.nombre as facultad, carr.nombre as carrera, 
                                    (select ma.codigo from sga_inscripcionmalla insmalla inner join sga_malla ma on insmalla.malla_id=ma.id where insmalla.inscripcion_id=ins.id and insmalla.status=True), 
                                    perso.telefono,ins.id as INSCRIPCION, 
                                    CASE WHEN perso.lgtbi = True THEN 'SI' else 'NO' END as lgtbi, 
                                    (select rasa.nombre from sga_perfilinscripcion perfil inner join sga_raza rasa on perfil.raza_id=rasa.id where perfil.persona_id=perso.id and perfil.status=True) as etnia, 
                                    perso.nacionalidad, pais.nombre as pais, provincia.nombre as provincia, canton.nombre as canton, 
                                    perso.direccion || ' ' || perso.direccion2 as direccion, 
                                    (select gru.codigo || ' ' || gru.nombre from socioecon_fichasocioeconomicainec fi inner join socioecon_gruposocioeconomico gru on fi.grupoeconomico_id=gru.id where fi.persona_id=perso.id and fi.status=True) as gruposocioeconomico, 
                                    ins.fechainicioprimernivel,ins.fechainicioconvalidacion, 
                                    (select CASE WHEN perfilin.tienediscapacidad = True THEN 'SI' else 'NO' END  from sga_perfilinscripcion perfilin where perfilin.persona_id=perso.id and perfilin.status=True) as tienediscapacidad, 
                                    (select disca.nombre  from sga_perfilinscripcion perfilin,sga_discapacidad disca where perfilin.persona_id=perso.id and perfilin.tipodiscapacidad_id=disca.id and perfilin.status=True) as discapacidad, 
                                    matri.id as idmatricula, CASE WHEN gruso.tipomatricula=1 THEN 'REGULAR' WHEN gruso.tipomatricula=2 THEN 'IRREGULAR' END as tipomatricula, 
                                    tima.nombre as tipoestudiante, ext.contactoemergencia as contactoemergencia, 
                                    uni.nombre as unidadeducativa, modal.nombre as modalidad,
                                    (select count(*) from sga_matricula matr where matr.inscripcion_id=ins.id and matr.status=True
                                    and matr.retiradomatricula=False) as veces_matricula , perso.ppl, matri.termino, mateasig.matriculas, 
                                    ins.estado_gratuidad
                                    from sga_matricula matri
                                    inner join sga_tipomatricula tima on tima.id=matri.tipomatricula_id and tima.status=True
                                    inner join sga_nivel ni on matri.nivel_id=ni.id and ni.status=True
                                    inner join sga_inscripcion ins on matri.inscripcion_id=ins.id and ins.status=True
	                                LEFT JOIN sga_materiaasignada mateasig ON mateasig.matricula_id = matri.id
                                    left join sga_coordinacion coor on coor.id=ins.coordinacion_id
                                    left join sga_carrera carr on carr.id=ins.carrera_id
                                    left join sga_institucionescolegio uni on uni.id=ins.unidadeducativa_id and uni.status=True
                                    left join sga_matriculagruposocioeconomico gruso on gruso.matricula_id=matri.id and gruso.status=True
                                    inner join sga_persona perso on ins.persona_id=perso.id
                                    left join med_personaextension ext on ext.persona_id=perso.id and ext.status=True
                                    left join sga_sexo se on se.id=perso.sexo_id and se.status=True
                                    left join sga_pais pais on perso.pais_id=pais.id and pais.status=True
                                    left join sga_provincia provincia on perso.provincia_id=provincia.id and provincia.status=True
                                    left join sga_canton canton on perso.canton_id=canton.id and canton.status=True
                                    inner join sga_periodo peri on ni.periodo_id=peri.id
                                    inner join sga_nivelmalla nimalla on matri.nivelmalla_id=nimalla.id 
                                    left join sga_sesion sesion on ins.sesion_id=sesion.id
                                    inner join sga_modalidad modal on ins.modalidad_id=modal.id
                                    where  matri.status=True and matri.retiradomatricula=False
                                    and ni.periodo_id= %s 
                                    and matri.id not in( 
                                    SELECT DISTINCT sga_matricula.id 
                                    FROM sga_matricula 
                                    INNER JOIN sga_nivel ON (sga_matricula.nivel_id = sga_nivel.id)   
                                    INNER JOIN sga_periodo ON (sga_nivel.periodo_id = sga_periodo.id)   
                                    INNER JOIN sga_materiaasignada ON (sga_matricula.id = sga_materiaasignada.matricula_id)  
                                    INNER JOIN sga_materia ON (sga_materiaasignada.materia_id = sga_materia.id)  
                                    INNER JOIN sga_asignatura ON (sga_materia.asignatura_id = sga_asignatura.id)  
                                    INNER JOIN sga_inscripcion ON (sga_matricula.inscripcion_id = sga_inscripcion.id)  
                                    WHERE (sga_nivel.periodo_id = %s AND sga_matricula.status = TRUE AND sga_asignatura.modulo = TRUE 
                                    AND ((SELECT COUNT(mta.id) 
                                    FROM sga_materiaasignada mta 
                                    WHERE mta.matricula_id=sga_matricula.id) = 1)  
                                    AND NOT (sga_inscripcion.carrera_id IN ( 
                                    SELECT U3.carrera_id AS Col1 
                                    FROM sga_coordinacion_carrera U3 
                                    WHERE U3.coordinacion_id = 9))  
                                    AND NOT (sga_inscripcion.carrera_id = 7)  
                                    AND NOT (sga_matricula.retiradomatricula = TRUE)))
                    """ % (periodoid, periodoid)

                    ###sga_matricula.estado_matricula = 2 AND
                    cursor.execute(listaestudiante)
                    results = cursor.fetchall()
                    # a = 0

                    for per in results:
                        a += 1
                        ws.write(a, 0, a - 4)

                        # ws.write(a, 1, '%s' % per[0])
                        # ws.write(a, 2, '%s' % per[1])
                        # ws.write(a, 3, '%s' % per[1])
                        # ws.write(a, 4, '%s' % per[2])

                        ws.write(a, 1, '%s' % per[3])
                        ws.write(a, 2, '%s' % per[4])
                        ws.write(a, 3, '%s' % per[5])
                        ws.write(a, 4, '%s' % per[11])
                        ws.write(a, 5, '%s' % per[32])
                        ws.write(a, 6, '%s' % per[36])

                        ws.write(a, 7, '%s' % per[6])
                        ws.write(a, 8, per[7], date_format)
                        ws.write(a, 9, '%s' % per[8])

                        # ws.write(a, 11, '%s' % per[9])
                        # ws.write(a, 12, '%s' % per[10])
                        # ws.write(a, 14, '%s' % per[12])

                        ws.write(a, 10, '%s' % per[13])
                        ws.write(a, 11, '%s' % per[25])
                        ws.write(a, 12, '%s' % per[26])

                        # ws.write(a, 16, '%s' % per[14])

                        ws.write(a, 13, '%s' % per[15])
                        ws.write(a, 14, '%s' % per[16])
                        ws.write(a, 15, '%s' % "SI" if per[34] else "NO")

                        ws.write(a, 16, '%s' % per[17])
                        ws.write(a, 17, '%s' % per[18])
                        ws.write(a, 18, '%s' % per[19])
                        ws.write(a, 19, '%s' % per[20])
                        ws.write(a, 20, '%s' % per[21])
                        ws.write(a, 21, '%s' % per[22])

                        # ws.write(a, 26, 'SI')
                        # ws.write(a, 27, '%s' % per[23])
                        # ws.write(a, 28, '%s' % per[24])
                        #
                        # ws.write(a, 31, '%s' % per[27])
                        # ws.write(a, 32, '%s' % per[28])
                        # ws.write(a, 33, '%s' % per[29])
                        ws.write(a, 22, '%s' % per[30])
                        ws.write(a, 23, '%s' % per[31])
                        if per[37] == 1:
                            gratu = 'GRATUIDAD COMPLETA'
                        elif per[37] == 2:
                            gratu = 'GRATUIDAD PARCIAL'
                        else:
                            gratu = 'PERDIDA DE GRATUIDA'

                        ws.write(a, 24, '%s' % gratu)

                    wb.save(response)
                    return response
                except Exception as ex:
                    pass




            elif action == 'reportehorarioexamen':
                try:
                    periodo = request.GET['periodo']

                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=horarioexamen.xls'

                    __author__ = 'Unemi'
                    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on',
                                      num_format_str='#,##0.00')
                    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                    title = easyxf(
                        'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    style1 = easyxf(num_format_str='D-MMM-YY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False



                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('exp_xls_post_part')
                    fmt = xlwt.easyxf
                    encabezado = fmt('font: height 150, bold on; border: left thin, right thin, top thin, bottom thin; align: wrap on, vert centre, horiz center;')

                    ws.write_merge(0, 0, 0, 6, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)

                    ws.col(0).width = 6000
                    ws.col(1).width = 6000
                    ws.col(2).width = 6000
                    ws.col(3).width = 6000
                    ws.col(4).width = 6000
                    ws.col(5).width = 6000
                    ws.col(6).width = 6000
                    ws.col(7).width = 6000
                    ws.col(8).width = 6000
                    ws.col(9).width = 6000


                    ws.write(4, 0, 'CARRERA', encabezado)
                    ws.write(4, 1, 'MODALIDAD', encabezado)

                    ws.write(4, 2, 'FACULTAD', encabezado)
                    # ws.write(4, 2, 'NIVEL')
                    ws.write(4, 3, 'PARALELO', encabezado)
                    ws.write(4, 4, 'ASIGNATURA', encabezado)
                    ws.write(4, 5, 'DOCENTE', encabezado)
                    ws.write(4, 6, 'FECHA', encabezado)
                    ws.write(4, 7, 'HORA', encabezado)
                    # ws.write(4, 8, 'PARCIAL')
                    ws.write(4, 8, 'BLOQUE-AULA', encabezado)

                    a = 4
                    date_format = xlwt.XFStyle()
                    date_format.num_format_str = 'yyyy/mm/dd'
                    horarioexamenes = HorarioExamen.objects.filter(materia__nivel__periodo=periodo,
                                                                   status=True).distinct().order_by(
                        'materia__asignaturamalla__malla__carrera__coordinacion',
                        'materia__asignaturamalla__nivelmalla', 'materia__asignaturamalla__malla__carrera',
                        'materia__paralelo')
                    for horarioexamen in horarioexamenes:
                        a += 1
                        ws.write(a, 0, u'%s' % horarioexamen.materia.asignaturamalla.malla.carrera.nombre)
                        ws.write(a, 1, u'%s' % horarioexamen.materia.nivel.modalidad.nombre)

                        ws.write(a, 2,
                                 u'%s' % horarioexamen.materia.asignaturamalla.malla.carrera.coordinacion_set.all()[
                                     0].nombre)
                        # ws.write(a, 2, u'%s' % horarioexamen.materia.asignaturamalla.nivelmalla.nombre)
                        ws.write(a, 3, u'%s' % horarioexamen.materia.paralelo)
                        ws.write(a, 4, u'%s' % horarioexamen.materia.asignatura.nombre)
                        ws.write(a, 5,
                                 u'%s' % horarioexamen.materia.profesor_principal().persona.nombre_completo_inverso() if horarioexamen.materia.profesor_principal() else '')
                        ws.write(a, 6, u'%s' % horarioexamen.fecha)
                        ws.write(a, 7, u'[%s a %s]' % (horarioexamen.turno.comienza.strftime("%H:%M %p"),
                            horarioexamen.turno.termina.strftime("%H:%M %p")))
                        # ws.write(a, 8, u'%s' % horarioexamen.detallemodelo.parcial())


                        aulas = ""
                        for aula in horarioexamen.materia.aulas():
                            aulas = aulas + aula.nombre
                            aulas = aulas + ","
                        if aulas.__len__() > 0:
                            aulas = aulas[0:aulas.__len__() - 1]
                        ws.write(a, 8, u'%s' % aulas)

                    a += 1
                    # ws.write_merge(a + 2, a + 2, 0, 1, datetime.today(), date_format)
                    wb.save(response)
                    return response
                except Exception as ex:
                    pass


            elif action == 'sinhorarios':
                try:
                    periodo = request.GET['periodo']

                    output = io.BytesIO()
                    workbook = xlsxwriter.Workbook(output)
                    ws = workbook.add_worksheet('Listado')

                    merge_format = workbook.add_format({
                        'bold': 1,
                        'border': 1,
                        'align': 'center',
                        'valign': 'vcenter',
                        'bg_color': 'silver',
                        'text_wrap': 1})

                    font_style2 = workbook.add_format({
                        'border': 1
                    })

                    ws.merge_range(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO',
                                   workbook.add_format({'align': 'center',
                                                        'valign': 'vcenter',
                                                        'bold': 1,
                                                        'font_size': 16}))

                    wb = Workbook(encoding='utf-8')

                    columns = [
                        (u"CARRERA", 20),
                        (u"MODALIDAD", 20),
                        (u"FACULTAD", 20),
                        (u"PARALELO", 20),
                        (u"MATERIA", 20)

                    ]
                    row_num = 3
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], merge_format)
                        ws.set_column(col_num, col_num, columns[col_num][1])

                    cursor = connections['sga_select'].cursor()
                    sql = "select a.nombre, m.paralelo, n.paralelo, nm.nombre, car.nombre, c.turno_id, coordinacion.nombre from sga_materia m " \
                          "inner join sga_nivel n on m.nivel_id=n.id and n.periodo_id=" + periodo + " " \
                                                                                                    "left join sga_clase c on c.materia_id=m.id " \
                                                                                                    "inner join sga_asignatura a on a.id=m.asignatura_id " \
                                                                                                    "inner join sga_asignaturamalla am on am.id=m.asignaturamalla_id " \
                                                                                                    "inner join sga_nivelmalla nm on nm.id=am.nivelmalla_id " \
                                                                                                    "inner join sga_malla malla on malla.id=am.malla_id " \
                                                                                                    "inner join sga_carrera car on car.id=malla.carrera_id " \
                                                                                                    "left JOIN sga_coordinacion_carrera coordcarr ON coordcarr.carrera_id=car.id "\
                                                                                                    "left JOIN sga_coordinacion coordinacion ON coordinacion.id=coordcarr.coordinacion_id " \
                                                                                                    "where c.turno_id is null "
                    cursor.execute(sql)
                    results = cursor.fetchall()
                    row_num = 4
                    for r in results:
                        i = 0
                        campo1 = r[0]
                        campo2 = r[1]
                        campo3 = r[2]
                        campo4 = r[3]
                        campo5 = r[4]
                        campo6 = r[6]
                        ws.write(row_num, 0, campo5, font_style2)
                        ws.write(row_num, 1, campo3, font_style2)

                        ws.write(row_num, 2, campo6, font_style2)

                        ws.write(row_num, 3, campo2, font_style2)
                        ws.write(row_num, 4, campo1, font_style2)



                        row_num += 1

                    workbook.close()
                    output.seek(0)
                    # Set up the Http response.
                    filename = 'DISTRIBUTIVO ASIGNATURA' + random.randint(1, 10000).__str__() + '.xlsx'
                    response = HttpResponse(output,
                                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = 'attachment; filename=%s' % filename

                    return response
                except Exception as ex:
                    pass


            elif action == 'avance_asistencia':
                try:
                    periodo = Periodo.objects.get(pk=int(request.POST['idperiodo']))

                    coordinaciones = Coordinacion.objects.filter(
                        carrera__malla__asignaturamalla__materia__nivel__periodo=periodo).order_by('-id').distinct()
                    diasnolaborables = DiasNoLaborable.objects.filter(periodo=periodo).order_by('fecha')
                    # claseshorarios = Clase.objects.filter(materia__nivel__periodo=periodo, activo=True).distinct().order_by('materia__profesormateria__profesor', 'turno__comienza')
                    output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media', 'asistencias'))
                    __author__ = 'Unemi'
                    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on',
                                      num_format_str='#,##0.00')
                    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                    title = easyxf(
                        'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    style1 = easyxf(num_format_str='D-MMM-YY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    for coordinacion1 in coordinaciones:
                        ws = wb.add_sheet(coordinacion1.alias)
                        coordinacionid = coordinacion1.id
                        ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                        ws.write(1, 0, "Periodo: " + periodo.nombre, font_style2)
                        ws.write(2, 0, "Fecha Corte: " + request.POST['fecha'], font_style2)
                        nombre = "AVANCE_" + datetime.now().strftime('%Y%m%d_%H%M%S') + ".xls"
                        filename = os.path.join(output_folder, nombre)
                        ruta = "media/asistencias/" + nombre
                        book = xlwt.Workbook()
                        columns = [
                            (u"FACULTAD", 6000),
                            (u"CARRERA", 6000),
                            (u"NIVEL", 6000),
                            (u"PARALELO", 6000),
                            (u"DOCENTE", 6000),
                            (u"ASIGNATURA", 6000),
                            (u"HORAS PROGRAMADAS MENSUAL", 6000),
                            (u"HORAS INGRESADAS", 6000),
                            (u"HORAS FALTAS", 6000),
                            (u"HORAS TOTAL ASISTENCIAS", 6000),
                            (u"PORCENTAJE AVANCE FECHA TOPE", 6000),
                            (u"PORCENTAJE AVANCE SEMESTRE", 6000),
                            (u"FECHA INICIO MATERIA", 6000),
                            (u"FECHA FIN MATERIA", 6000),
                        ]
                        row_num = 3
                        for col_num in range(len(columns)):
                            ws.write(row_num, col_num, columns[col_num][0], font_style)
                            ws.col(col_num).width = columns[col_num][1]

                        # profesormateria = ProfesorMateria.objects.filter(materia__nivel__periodo=periodo, principal=True).distinct().order_by('materia__asignaturamalla__malla__carrera','profesor')
                        # asistencialeccion = AsistenciaLeccion.objects.filter(materiaasignada__materia__nivel__periodo=periodo, leccion__fecha__lte=fecha).order_by("materiaasignada__materia").distinct("materiaasignada__materia")
                        cursor = connections['sga_select'].cursor()
                        sql = "select tabladocente.facultad , tabladocente.carrera, tabladocente.nivel, tabladocente.paralelo, tabladocente.docente,tabladocente.asignatura, tabladocente.horas , COALESCE(tablaasistencia.asistencia,0) as asistencia, tabladocente.inicio, tabladocente.fin,  " \
                              " (select count(*) from sga_faltasmateriaperiodo fm1 where fm1.materia_id=tabladocente.materiaid) as faltas " \
                              " from (select pm.id as profesormateriaid, m.inicio, m.fin, m.id as materiaid, co.nombre as facultad, ca.nombre as carrera, nm.nombre as nivel, m.paralelo, (per.apellido1 || ' ' || per.apellido2 || ' ' || per.nombres) as docente,asi.nombre as asignatura, m.horas " \
                              " from sga_profesormateria pm, sga_materia m, sga_nivel n, sga_asignaturamalla am, sga_malla ma,sga_carrera ca, sga_coordinacion_carrera cc, sga_coordinacion co, sga_nivelmalla nm, sga_profesor pr, sga_persona per, sga_asignatura asi " \
                              " where m.id=pm.materia_id and pm.tipoprofesor_id not in (4) and n.id=m.nivel_id and am.id=m.asignaturamalla_id and ma.id=am.malla_id and ca.id=ma.carrera_id and cc.carrera_id=ca.id and co.id=cc.coordinacion_id and nm.id=am.nivelmalla_id and pr.id=pm.profesor_id and per.id=pr.persona_id " \
                              " and asi.id=m.asignatura_id and n.periodo_id= " + request.POST[
                                  'idperiodo'] + " and co.id=" + str(
                            coordinacionid) + " order by co.nombre, ca.nombre, nm.nombre, m.paralelo, docente) as tabladocente " \
                                              " left join (select mat1.id as materiaid, count(mat1.id) as asistencia from sga_leccion l1 , sga_clase c1 , sga_materia mat1, sga_nivel ni1 where l1.clase_id=c1.id and c1.materia_id=mat1.id and mat1.nivel_id=ni1.id and l1.fecha<= '" + \
                              request.POST['fecha'] + "' and ni1.periodo_id=" + request.POST['idperiodo'] + " " \
                                                                                                            " and l1.fecha not in (select dnl1.fecha from sga_diasnolaborable dnl1 where dnl1.periodo_id=" + \
                              request.POST[
                                  'idperiodo'] + ") GROUP by mat1.id) as tablaasistencia on tablaasistencia.materiaid=tabladocente.materiaid order by tabladocente.facultad , tabladocente.carrera, tabladocente.nivel, tabladocente.paralelo, tabladocente.docente"
                        cursor.execute(sql)
                        results = cursor.fetchall()
                        row_num = 4
                        for r in results:
                            i = 0
                            campo1 = r[0]
                            campo2 = r[1]
                            campo3 = r[2]
                            campo4 = r[3]
                            campo5 = r[4]
                            campo6 = r[5]
                            campo7 = r[6]
                            campo8 = r[7]
                            campo9 = 0
                            if r[6] > 0:
                                porcentajeperiodo = round((float(r[7]) / r[6]) * 100, 2)
                                if porcentajeperiodo > 100:
                                    porcentajeperiodo = 100
                                campo9 = porcentajeperiodo
                            campo13 = 0
                            if (r[10] + r[7]) > 0:
                                porcentajefecha = round((float(r[7]) / (r[10] + r[7])) * 100, 2)
                                if porcentajefecha > 100:
                                    porcentajefecha = 100
                                campo13 = porcentajefecha

                            campo10 = r[8]
                            campo11 = r[9]
                            campo12 = r[10]

                            ws.write(row_num, 0, campo1, font_style2)
                            ws.write(row_num, 1, campo2, font_style2)
                            ws.write(row_num, 2, campo3, font_style2)
                            ws.write(row_num, 3, campo4, font_style2)
                            ws.write(row_num, 4, campo5, font_style2)
                            ws.write(row_num, 5, campo6, font_style2)
                            ws.write(row_num, 6, campo7, font_style2)
                            ws.write(row_num, 7, campo8, font_style2)
                            ws.write(row_num, 8, campo12, font_style2)
                            ws.write(row_num, 9, campo12 + campo8, font_style2)
                            ws.write(row_num, 10, campo13, font_style2)
                            ws.write(row_num, 11, campo9, font_style2)
                            ws.write(row_num, 12, campo10, style1)
                            ws.write(row_num, 13, campo11, style1)
                            # while i < len(r):
                            #     # ws.write(row_num, i, r[i], font_style)
                            #     # ws.col(i).width = columns[i][1]
                            row_num += 1
                    wb.save(filename)
                    # return book
                    return JsonResponse({'result': 'ok', 'archivo': ruta})
                except Exception as ex:
                    pass


            elif action == 'reportealumnosinprofesor':
                try:
                    output = io.BytesIO()
                    workbook = xlsxwriter.Workbook(output)
                    ws = workbook.add_worksheet('exp_xls_post_part')

                    merge_format = workbook.add_format({
                        'bold': 1,
                        'border': 1,
                        'align': 'center',
                        'valign': 'vcenter'})

                    ws.merge_range(0, 0, 0, 7, 'UNIVERSIDAD ESTATAL DE MILAGRO',
                                   workbook.add_format({'align': 'center', 'valign': 'vcenter'}))
                    columns = [

                        (u"CEDULA", 25),
                        (u"PRIMER APELLIDO", 40),
                        (u"SEGUNDO APELLIDO", 40),
                        (u"NOMBRES", 40),
                        (u"FACULTAD", 40),
                        (u"CARRERA", 40),
                        (u"MODALIDAD", 25),
                        (u"MATERIA", 40),
                        (u"PARALELO", 25),
                        (u"EMAIL", 40),
                        # 7(u"EMAIL INSTITUCIONAL", 40),
                        (u"CELULAR", 25),
                        (u"TELEFONO", 25)

                        # 12(u"NIVEL", 25),
                        # 13(u"NOTA FINAL", 40),
                        # 14(u"ASISTENCIA FINAL", 40),
                        # 15(u"ESTADO", 25),
                        # 16(u"NUMERO DE MATRICULA", 25),
                        # 17(u"ANIO MALLA", 25),
                        # 18(u"FECHA NACIMIENTO", 25),

                        # 20(u"SEXO", 25),
                        # 21(u"CANTON RESIDENCIA", 25),
                        # 22(u"GRUPO SOCIECONOMICO", 25)
                    ]
                    row_num = 3
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], merge_format)
                        ws.set_column(col_num, col_num, columns[col_num][1])
                    cursor = connections['default'].cursor()
                    sql = """select coordinacion.nombre as coordinacion, CASE WHEN carrera.mencion = '' THEN carrera.nombre ELSE (carrera.nombre || ' CON MENCION EN ' || carrera.mencion) END as carrera, 
                                       persona.cedula, persona.apellido1, persona.apellido2, persona.nombres, persona.email, persona.emailinst, persona.telefono, persona.telefono_conv, 
                                       (asignatura.nombre || CASE WHEN materia.alias = '' THEN '' ELSE (' - ('|| materia.alias || ')') END  || CASE WHEN materia.identificacion = '' THEN '' ELSE (' - ['|| materia.identificacion || ']') END) as materia, 
                                       materia.paralelo as paralelo, nivelmalla.nombre as nivel, materiaasig.notafinal as notafinal, materiaasig.asistenciafinal as asistenciafinal, tipoestado.nombre as estado,  materiaasig.matriculas as veces, extract(year from malla.inicio) as aniomalla,  
                                       persona.nacimiento, modalidad.nombre, genero.nombre, canton.nombre, gruposocio.nombre  
                                       from sga_materiaasignada materiaasig 
                                       inner join sga_materia materia on materia.id = materiaasig.materia_id 
                                       inner join sga_nivel nivel on nivel.id = materia.nivel_id 
                                       inner join sga_modalidad modalidad on nivel.modalidad_id = modalidad.id 
                                       inner join sga_asignaturamalla asigmalla on asigmalla.id = materia.asignaturamalla_id 
                                       inner join sga_malla malla on malla.id = asigmalla.malla_id 
                                       inner join sga_carrera carrera on carrera.id = malla.carrera_id 
                                       inner join sga_matricula matricula on matricula.id = materiaasig.matricula_id 
                                       inner join sga_inscripcion inscripcion on inscripcion.id = matricula.inscripcion_id 
                                       inner join sga_persona persona on persona.id = inscripcion.persona_id 
                                       left join sga_sexo genero on genero.id = persona.sexo_id 
                                       left join sga_canton canton on canton.id = persona.canton_id 
                                       inner join sga_asignatura asignatura on asignatura.id = materia.asignatura_id 
                                       inner join sga_nivelmalla nivelmalla on nivelmalla.id = asigmalla.nivelmalla_id 
                                       inner join sga_tipoestado tipoestado on tipoestado.id = materiaasig.estado_id 
                                       inner join sga_coordinacion_carrera coordinacioncarrera on coordinacioncarrera.carrera_id = carrera.id 
                                       inner join sga_coordinacion coordinacion on coordinacion.id = coordinacioncarrera.coordinacion_id 
                                       inner join sga_matriculagruposocioeconomico matri_gruposocio on matri_gruposocio.matricula_id = matricula.id
                                       inner join socioecon_gruposocioeconomico gruposocio on matri_gruposocio.gruposocioeconomico_id = gruposocio.id
                                       where coordinacion.id not in(9) and nivel.periodo_id = %s""" % periodo.id
                    cursor.execute(sql)
                    results = cursor.fetchall()
                    row_num = 4
                    for r in results:
                        ws.write(row_num, 0, u'%s' % r[2])
                        ws.write(row_num, 1, u'%s' % r[3])
                        ws.write(row_num, 2, u'%s' % r[4])
                        ws.write(row_num, 3, u'%s' % r[5])

                        ws.write(row_num, 4, u'%s' % r[0])
                        ws.write(row_num, 5, u'%s' % r[1])
                        ws.write(row_num, 6, u'%s' % r[19])
                        ws.write(row_num, 7, u'%s' % r[10])
                        ws.write(row_num, 8, u'%s' % r[11])
                        ws.write(row_num, 9, u'%s' % r[6])
                        ws.write(row_num, 10, u'%s' % r[8])
                        ws.write(row_num, 11, u'%s' % r[9])
                        row_num += 1
                    workbook.close()
                    output.seek(0)
                    # Set up the Http response.
                    filename = 'reporte_asignaturas_estudiantes_sin_profesor' + random.randint(1,
                                                                                               10000).__str__() + '.xlsx'
                    response = HttpResponse(output,
                                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = 'attachment; filename=%s' % filename

                    return response
                except Exception as ex:
                    pass

            elif action == 'reportedistributivo':
                try:

                    output = io.BytesIO()
                    workbook = xlsxwriter.Workbook(output)
                    ws = workbook.add_worksheet('Listado')

                    merge_format = workbook.add_format({
                        'bold': 1,
                        'border': 1,
                        'align': 'center',
                        'valign': 'vcenter',
                        'bg_color': 'silver',
                        'text_wrap': 1})

                    font_style2 = workbook.add_format({
                        'border': 1
                    })

                    ws.merge_range(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO',
                                   workbook.add_format({'align': 'center',
                                                        'valign': 'vcenter',
                                                        'bold': 1,
                                                        'font_size': 16}))

                    wb = Workbook(encoding='utf-8')

                    columns = [
                        (u"PERIODO", 20),
                        (u"FACULTAD", 20),
                        (u"CARRERA", 20),
                        (u"MALLA", 20),
                        (u"MODELO EVALUATIVO", 20),
                        (u"JORNADA", 20),
                        # 5(u"NIVEL", 6000),
                        (u"PARALELO", 20),
                        (u"ASIGNATURA", 20),
                        (u"INICIO MATERIA", 20),
                        (u"FIN MATERIA", 20),
                        # 8(u"TEORICA PRACTICA", 6000),
                        (u"CAPACIDAD", 20),
                        (u"MATRICULADOS", 20),
                        (u"DOCENTE", 20),
                        (u"CEDULA", 20),
                        # (u"USUARIO", 6000),
                       # 13 (u"AFINIDAD", 6000),
                        (u"HORAS SEMANALES", 20),
                        (u"MALLA (HORAS PRESENCIALES SEMANALES)", 20),
                        (u"TIPO", 20),
                        (u"CORREO PERSONAL", 20),
                        (u"CORREO INSTITUCIONAL", 20),
                        (u"TIPO PROFESOR", 20),
                        (u"PROFESOR DESDE", 20),
                        (u"PROFESOR HASTA", 20),
                        (u"DEDICACION", 20),
                        (u"CATEGORIA", 20),
                        # 26(u"FIN ASISTENCIA", 4000),
                        # (u"ID", 4000),
                        (u"TELEFONO", 20),
                        # 29(u"IDMATERIA", 2500),
                        # 30(u"ACEPTACION", 2500),
                        # 31(u"OBSERVACION ACEPTACION", 10000),
                        # 32(u"HORARIO FECHA ACEPTACION", 10000),
                        # 33(u"HORARIO ACEPTACION", 10000),
                        # 34(u"HORARIO OBSERVACION ACEPTACION", 10000),
                        # 35(u"IDMOODLE", 2500),
                        (u"PAIS DE RECIDENCIA", 20),
                        (u"PROVINCIA", 20),
                        (u"CIUDAD", 20),
                        (u"NACIONALIDAD", 20),
                        (u"ETNIA", 20),
                        (u"GENERO", 20),
                        # 42(u"PERIODO", 6000),

                    ]
                    periodo = request.GET['periodo']
                    row_num = 3
                    date_format = xlwt.XFStyle()
                    date_format.num_format_str = 'yyyy/mm/dd'
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], merge_format)
                        ws.set_column(col_num, col_num, columns[col_num][1])

                    cursor = connections['sga_select'].cursor()
                    # sql = "SELECT " \
                    #       " sga_coordinacion.nombre AS Facultad, " \
                    #       " sga_carrera.nombre AS Carrera, " \
                    #       " sga_sesion.nombre AS Seccion, " \
                    #       " sga_nivelmalla.nombre AS Nivel, " \
                    #       " sga_materia.paralelo as Paralelo, " \
                    #       " sga_materia.id AS Idmateria, " \
                    #       " sga_asignatura.nombre AS Asignatura," \
                    #       " sga_persona.apellido1 || ' ' || sga_persona.apellido2 || ' ' || sga_persona.nombres  AS Docente," \
                    #       " sga_profesormateria.hora AS sga_profesormateria_hora, " \
                    #       " (case sga_profesormateria.principal when true then 'PRINCIPAL' else 'PRACTICA' end) as Tipo," \
                    #       " sga_persona.cedula, (select u.username from auth_user u where u.id=sga_persona.usuario_id), sga_persona.email, sga_persona.emailinst," \
                    #       " sga_materia.cupo as cupo, (select count(*) from sga_materiaasignada ma, sga_matricula mat1 where ma.matricula_id=mat1.id and mat1.estado_matricula in (2,3) and ma.materia_id=sga_materia.id and ma.id not in (select mr.materiaasignada_id from sga_materiaasignadaretiro mr)) as nmatriculados," \
                    #       " (select count(*) from sga_materiaasignada ma, sga_matricula mat1 where ma.matricula_id=mat1.id and mat1.termino=True and ma.materia_id=sga_materia.id and ma.id not in (select mr.materiaasignada_id from sga_materiaasignadaretiro mr)) as nmatriculados_acpta_termino," \
                    #       " sga_tipoprofesor.nombre as Tipoprofesor, " \
                    #       " sga_profesormateria.desde as desde, " \
                    #       " sga_profesormateria.hasta as hasta, " \
                    #       " (select ti.nombre from sga_profesordistributivohoras dis,sga_tiempodedicaciondocente ti where dis.dedicacion_id=ti.id and dis.profesor_id=sga_profesor.id and periodo_id=" + periodo + " and dis.status=True) as dedicacion," \
                    #                                                                                                                                                                                                " (select ca.nombre from sga_profesordistributivohoras dis,sga_categorizaciondocente ca where dis.categoria_id=ca.id and dis.profesor_id=sga_profesor.id and dis.periodo_id=" + periodo + " and dis.status=True) as categoria, " \
                    #                                                                                                                                                                                                                                                                                                                                                                                          " (case sga_asignaturamalla.practicas when true then 'SI' else 'NO' end) as tipomateria, " \
                    #                                                                                                                                                                                                                                                                                                                                                                                          " (case sga_profesormateria.afinidad when true then 'SI' else 'NO' end) as afinidad, " \
                    #                                                                                                                                                                                                                                                                                                                                                                                          " sga_materia.inicio as inicio, " \
                    #                                                                                                                                                                                                                                                                                                                                                                                          " sga_materia.fin as fin, " \
                    #                                                                                                                                                                                                                                                                                                                                                                                          " sga_materia.fechafinasistencias as finasistencia," \
                    #                                                                                                                                                                                                                                                                                                                                                                                          " sga_materia.id as id, sga_persona.telefono_conv as telefonoconv, sga_persona.telefono as telefono, extract(year from sga_malla.inicio) as anio," \
                    #                                                                                                                                                                                                                                                                                                                                                                                          " (select modelo.nombre from sga_modeloevaluativo modelo where modelo.id = sga_materia.modeloevaluativo_id) as modeloevaluativo," \
                    #                                                                                                                                                                                                                                                                                                                                                                                          " sga_asignaturamalla.horaspresencialessemanales as horaspresencialessemanales," \
                    #                                                                                                                                                                                                                                                                                                                                                                                          " sga_profesormateria.aceptarmateria as aceptarmateria," \
                    #                                                                                                                                                                                                                                                                                                                                                                                          " sga_profesormateria.aceptarmateriaobs as aceptarmateriaobs, " \
                    #                                                                                                                                                                                                                                                                                                                                                                                          " sga_profesormateria.fecha_horario as fecha_horario, " \
                    #                                                                                                                                                                                                                                                                                                                                                                                          " sga_profesormateria.aceptarhorario as aceptarhorario, " \
                    #                                                                                                                                                                                                                                                                                                                                                                                          " sga_profesormateria.aceptarhorarioobs as aceptarhorarioobs," \
                    #                                                                                                                                                                                                                                                                                                                                                                                          " sga_materia.idcursomoodle as idcursomoodle , prov.nombre AS provincia, cant.nombre AS ciudad " \
                    #                                                                                                                                                                                                                                                                                                                                                                                          " FROM public.sga_materia sga_materia" \
                    #                                                                                                                                                                                                                                                                                                                                                                                          " LEFT JOIN public.sga_profesormateria sga_profesormateria ON sga_materia.id = sga_profesormateria.materia_id and sga_profesormateria.status=true and sga_profesormateria.activo=true" \
                    #                                                                                                                                                                                                                                                                                                                                                                                          " LEFT JOIN public.sga_profesor sga_profesor ON sga_profesor.id = sga_profesormateria.profesor_id and sga_profesor.status=true " \
                    #                                                                                                                                                                                                                                                                                                                                                                                          " LEFT JOIN public.sga_tipoprofesor sga_tipoprofesor ON sga_tipoprofesor.id = sga_profesormateria.tipoprofesor_id and sga_tipoprofesor.status=true" \
                    #                                                                                                                                                                                                                                                                                                                                                                                          " LEFT JOIN public.sga_persona sga_persona ON sga_profesor.persona_id = sga_persona.id and sga_persona.status=true " \
                    #                                                                                                                                                                                                                                                                                                                                                                                          " LEFT JOIN sga_provincia prov ON prov.id=sga_persona.provincia_id " \
                    #                                                                                                                                                                                                                                                                                                                                                                                          " LEFT JOIN sga_canton cant ON cant.id=sga_persona.canton_id " \
                    #                                                                                                                                                                                                                                                                                                                                                                                          " INNER JOIN public.sga_nivel sga_nivel ON sga_materia.nivel_id = sga_nivel.id and sga_nivel.status=true " \
                    #                                                                                                                                                                                                                                                                                                                                                                                          " INNER JOIN public.sga_asignatura sga_asignatura ON sga_materia.asignatura_id = sga_asignatura.id and sga_asignatura.status=true" \
                    #                                                                                                                                                                                                                                                                                                                                                                                          " INNER JOIN public.sga_asignaturamalla sga_asignaturamalla ON sga_materia.asignaturamalla_id = sga_asignaturamalla.id and sga_asignaturamalla.status=true" \
                    #                                                                                                                                                                                                                                                                                                                                                                                          " INNER JOIN public.sga_nivelmalla sga_nivelmalla ON sga_asignaturamalla.nivelmalla_id = sga_nivelmalla.id and sga_nivelmalla.status=true " \
                    #                                                                                                                                                                                                                                                                                                                                                                                          " INNER JOIN public.sga_malla sga_malla ON sga_asignaturamalla.malla_id = sga_malla.id and sga_malla.status=true " \
                    #                                                                                                                                                                                                                                                                                                                                                                                          " INNER JOIN public.sga_carrera sga_carrera ON sga_malla.carrera_id = sga_carrera.id and sga_carrera.status=true  " \
                    #                                                                                                                                                                                                                                                                                                                                                                                          " INNER JOIN public.sga_coordinacion_carrera sga_coordinacion_carrera ON sga_carrera.id = sga_coordinacion_carrera.carrera_id" \
                    #                                                                                                                                                                                                                                                                                                                                                                                          " INNER JOIN public.sga_coordinacion sga_coordinacion ON sga_coordinacion_carrera.coordinacion_id = sga_coordinacion.id and sga_coordinacion.status=true" \
                    #                                                                                                                                                                                                                                                                                                                                                                                          " INNER JOIN public.sga_sesion sga_sesion ON sga_nivel.sesion_id = sga_sesion.id " \
                    #                                                                                                                                                                                                                                                                                                                                                                                          " INNER JOIN public.sga_periodo sga_periodo ON sga_nivel.periodo_id = sga_periodo.id" \
                    #                                                                                                                                                                                                                                                                                                                                                                                          " WHERE sga_periodo.id = " + periodo + " and sga_materia.status=true" \
                    #                                                                                                                                                                                                                                                                                                                                                                                                                                 " ORDER BY sga_coordinacion.nombre, sga_carrera.nombre, sga_sesion.nombre, sga_nivelmalla.nombre,sga_materia.paralelo,sga_asignatura.nombre"

                    sql= f"""
                        SELECT sga_coordinacion.nombre AS Facultad, sga_carrera.nombre AS Carrera, sga_sesion.nombre AS Seccion, sga_nivelmalla.nombre AS Nivel, sga_materia.paralelo AS Paralelo, sga_materia.id AS Idmateria, sga_asignatura.nombre AS Asignatura, sga_persona.apellido1 || ' ' || sga_persona.apellido2 || ' ' || sga_persona.nombres AS Docente, sga_profesormateria.hora AS sga_profesormateria_hora, (CASE sga_profesormateria.principal WHEN TRUE THEN 'PRINCIPAL' ELSE 'PRACTICA' END) AS Tipo, sga_persona.cedula, (
                            SELECT u.username
                            FROM auth_user u
                            WHERE u.id=sga_persona.usuario_id), sga_persona.email, sga_persona.emailinst, sga_materia.cupo AS cupo, (
                            SELECT COUNT(*)
                            FROM sga_materiaasignada ma, sga_matricula mat1
                            WHERE ma.matricula_id=mat1.id AND ma.status=True AND mat1.status=True AND mat1.estado_matricula in (2,3) AND ma.materia_id=sga_materia.id AND ma.id NOT in (
                            SELECT mr.materiaasignada_id
                            FROM sga_materiaasignadaretiro mr)) AS nmatriculados, sga_tipoprofesor.nombre AS Tipoprofesor,
                             sga_profesormateria.desde AS desde, sga_profesormateria.hasta AS hasta, (
                            SELECT ti.nombre
                            FROM sga_profesordistributivohoras dis,sga_tiempodedicaciondocente ti
                            WHERE dis.dedicacion_id=ti.id AND dis.profesor_id=sga_profesor.id AND periodo_id={periodo} AND dis.status= TRUE) AS dedicacion, (
                            SELECT ca.nombre
                            FROM sga_profesordistributivohoras dis,sga_categorizaciondocente ca
                            WHERE dis.categoria_id=ca.id AND dis.profesor_id=sga_profesor.id AND dis.periodo_id={periodo} AND dis.status= TRUE) AS categoria, 
                            (CASE sga_asignaturamalla.practicas WHEN TRUE THEN 'SI' ELSE 'NO' END) AS tipomateria, 
                            (CASE sga_profesormateria.afinidad WHEN TRUE THEN 'SI' ELSE 'NO' END) AS afinidad, 
                            sga_materia.inicio AS inicio, sga_materia.fin AS fin, sga_materia.fechafinasistencias AS finasistencia, sga_materia.id AS id, sga_persona.telefono_conv AS telefonoconv, sga_persona.telefono AS telefono, EXTRACT(YEAR
                            FROM sga_malla.inicio) AS anio, (
                            SELECT modelo.nombre
                            FROM sga_modeloevaluativo modelo
                            WHERE modelo.id = sga_materia.modeloevaluativo_id) AS modeloevaluativo, sga_asignaturamalla.horaspresencialessemanales AS horaspresencialessemanales, sga_profesormateria.aceptarmateria AS aceptarmateria, sga_profesormateria.aceptarmateriaobs AS aceptarmateriaobs, sga_profesormateria.fecha_horario AS fecha_horario, sga_profesormateria.aceptarhorario AS aceptarhorario, sga_profesormateria.aceptarhorarioobs AS aceptarhorarioobs, sga_materia.idcursomoodle AS idcursomoodle, prov.nombre AS provincia, cant.nombre AS ciudad,
                            pais.nombre as pais, sga_persona.nacionalidad, sex.nombre as genero, raza.nombre as etnia, sga_periodo.nombre as periodon
                            ,(
                            SELECT COUNT(*)
                            FROM sga_materiaasignada ma2, sga_matricula mat1
                            WHERE ma2.matricula_id=mat1.id AND mat1.termino= TRUE AND ma2.materia_id=sga_materia.id AND ma2.id NOT in (
                            SELECT mr.materiaasignada_id
                            FROM sga_materiaasignadaretiro mr)) AS nmatriculados_acpta_termino
                            FROM public.sga_materia sga_materia
                            LEFT JOIN public.sga_profesormateria sga_profesormateria ON sga_materia.id = sga_profesormateria.materia_id AND sga_profesormateria.status= TRUE AND sga_profesormateria.activo= TRUE
                            LEFT JOIN public.sga_profesor sga_profesor ON sga_profesor.id = sga_profesormateria.profesor_id AND sga_profesor.status= TRUE
                            LEFT JOIN public.sga_tipoprofesor sga_tipoprofesor ON sga_tipoprofesor.id = sga_profesormateria.tipoprofesor_id AND sga_tipoprofesor.status= TRUE
                            LEFT JOIN public.sga_persona sga_persona ON sga_profesor.persona_id = sga_persona.id AND sga_persona.status= TRUE
                            LEFT JOIN sga_provincia prov ON prov.id=sga_persona.provincia_id
                            LEFT JOIN sga_canton cant ON cant.id=sga_persona.canton_id
                            LEFT JOIN sga_perfilinscripcion perfil on perfil.persona_id = sga_persona.id
                            LEFT JOIN sga_sexo sex ON sex.id = sga_persona.sexo_id
                            LEFT JOIN sga_raza raza ON raza.id = perfil.raza_id
                            LEFT JOIN sga_pais pais ON pais.id = prov.pais_id
                            INNER JOIN public.sga_nivel sga_nivel ON sga_materia.nivel_id = sga_nivel.id AND sga_nivel.status= TRUE
                            INNER JOIN public.sga_asignatura sga_asignatura ON sga_materia.asignatura_id = sga_asignatura.id AND sga_asignatura.status= TRUE
                            INNER JOIN public.sga_asignaturamalla sga_asignaturamalla ON sga_materia.asignaturamalla_id = sga_asignaturamalla.id AND sga_asignaturamalla.status= TRUE
                            INNER JOIN public.sga_nivelmalla sga_nivelmalla ON sga_asignaturamalla.nivelmalla_id = sga_nivelmalla.id AND sga_nivelmalla.status= TRUE
                            INNER JOIN public.sga_malla sga_malla ON sga_asignaturamalla.malla_id = sga_malla.id AND sga_malla.status= TRUE
                            INNER JOIN public.sga_carrera sga_carrera ON sga_malla.carrera_id = sga_carrera.id AND sga_carrera.status= TRUE
                            INNER JOIN public.sga_coordinacion_carrera sga_coordinacion_carrera ON sga_carrera.id = sga_coordinacion_carrera.carrera_id
                            INNER JOIN public.sga_coordinacion sga_coordinacion ON sga_coordinacion_carrera.coordinacion_id = sga_coordinacion.id AND sga_coordinacion.status= TRUE
                            INNER JOIN public.sga_sesion sga_sesion ON sga_nivel.sesion_id = sga_sesion.id
                            INNER JOIN public.sga_periodo sga_periodo ON sga_nivel.periodo_id = sga_periodo.id
                            WHERE sga_periodo.id = {periodo} AND sga_materia.status= TRUE
                            ORDER BY sga_coordinacion.nombre, sga_carrera.nombre, sga_sesion.nombre, sga_nivelmalla.nombre,sga_materia.paralelo,sga_asignatura.nombre
                    """

                    cursor.execute(sql)
                    results = cursor.fetchall()
                    row_num = 4
                    for r in results:
                        i = 0
                        campo1 = r[0].__str__()
                        campo2 = r[1].__str__()
                        campo3 = r[2].__str__()
                        campo4 = r[3].__str__()
                        campo5 = r[4].__str__()
                        campo6 = r[5]
                        campo7 = r[6].__str__()
                        campo8 = r[7].__str__()
                        campo9 = r[8]
                        campo10 = r[9].__str__()
                        campo11 = r[10].__str__()
                        # campo12 = r[11]
                        campo13 = r[12].__str__()
                        campo14 = r[13].__str__()
                        campo15 = r[14].__str__()
                        campo16 = r[15].__str__()
                        campo17 = r[16].__str__()
                        campo18 = r[17].__str__()
                        campo19 = r[18].__str__()
                        campo20 = r[19].__str__()
                        campo21 = r[20].__str__()
                        campo22 = r[21].__str__()
                        campo23 = r[22].__str__()
                        campo24 = r[23].__str__()
                        campo25 = r[24].__str__()
                        campo26 = r[25].__str__()
                        # campo27 = r[26]
                        campo28 = r[27].__str__() + " - " + r[28].__str__()
                        campo29 = r[29].__str__()
                        campo30 = r[30].__str__()
                        campo31 = r[31].__str__()
                        if r[33] == None or r[33] == '':
                            campo32 = ''
                            campo33 = ''
                        else:
                            campo32 = 'NO'
                            if r[32]:
                                campo32 = 'SI'
                            campo33 = r[33]
                        campo34 = r[34]
                        if r[36] == None or r[36] == '':
                            campo35 = ''
                            campo36 = ''
                        else:
                            campo35 = 'NO'
                            if r[35]:
                                campo35 = 'SI'
                            campo36 = r[36]
                        campo37 = r[37]
                        campo38 = r[38]
                        campo39 = r[39]
                        campo_cant_acapta = r[40]
                        pais = r[40]
                        nacionalidad = r[41]
                        genero = r[42]
                        etnia = r[43]
                        periodon = r[44]
                        ws.write(row_num, 0, periodon, font_style2)
                        ws.write(row_num, 1, campo1, font_style2)
                        ws.write(row_num, 2, campo2, font_style2)
                        ws.write(row_num, 3, campo29, font_style2)
                        ws.write(row_num, 4, campo30, font_style2)

                        ws.write(row_num, 5, campo3, font_style2)

                        # ws.write(row_num, 4, campo4, font_style2)

                        ws.write(row_num, 6, campo5, font_style2)
                        ws.write(row_num, 7, campo7, font_style2)
                        ws.write(row_num, 8, campo24, font_style2)
                        ws.write(row_num, 9, campo25, font_style2)

                        # ws.write(row_num, 7, campo22, font_style2)

                        ws.write(row_num, 10, campo15, font_style2)
                        ws.write(row_num, 11, campo16, font_style2)
                        ws.write(row_num, 12, campo8, font_style2)
                        ws.write(row_num, 13, campo11, font_style2)
                        # ws.write(row_num, 12, campo12, font_style2)
                        # ws.write(row_num, 12, campo23, font_style2)
                        ws.write(row_num, 14, campo9, font_style2)
                        ws.write(row_num, 15, campo9, font_style2)
                        ws.write(row_num, 16, campo10, font_style2)
                        ws.write(row_num, 17, campo13, font_style2)
                        ws.write(row_num, 18, campo14, font_style2)
                        ws.write(row_num, 19, campo17, font_style2)
                        ws.write(row_num, 20, campo18, font_style2)
                        ws.write(row_num, 21, campo19, font_style2)
                        ws.write(row_num, 22, campo20, font_style2)
                        ws.write(row_num, 23, campo21, font_style2)
                        # ws.write(row_num, 25, campo26, style1)
                        # ws.write(row_num, 25, campo27, font_style2)
                        ws.write(row_num, 24, campo28, font_style2)

                        # ws.write(row_num, 28, campo6, font_style2)
                        # ws.write(row_num, 29, campo32, font_style2)
                        # ws.write(row_num, 30, campo33, font_style2)
                        # ws.write(row_num, 31, campo34, date_format)
                        # ws.write(row_num, 32, campo35, font_style2)
                        # ws.write(row_num, 33, campo36, font_style2)
                        # ws.write(row_num, 34, campo37, font_style2)
                        ws.write(row_num, 25, campo38, font_style2)
                        ws.write(row_num, 26, campo39, font_style2)
                        ws.write(row_num, 27, pais, font_style2)
                        ws.write(row_num, 28, nacionalidad, font_style2)
                        ws.write(row_num, 29, etnia, font_style2)
                        ws.write(row_num, 30, genero , font_style2)
                        row_num += 1

                    workbook.close()
                    output.seek(0)
                    # Set up the Http response.
                    filename = 'DISTRIBUTIVO ASIGNATURA' + random.randint(1, 10000).__str__() + '.xlsx'
                    response = HttpResponse(output,
                                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = 'attachment; filename=%s' % filename

                    return response

                except Exception as ex:
                    pass

            elif action == 'listado_ppl':
                try:
                    __author__ = 'Unemi'
                    inscripciones = Inscripcion.objects.filter(status=True, carrera__coordinacion__id=9, persona__ppl=True).distinct().order_by( 'persona__apellido1', 'persona__apellido2','persona__nombres')
                    #matriculas = Matricula.objects.filter(status=True,nivel__periodo=periodo, inscripcion__modalidad__id__in = [1,2], inscripcion__carrera__coordinacion__id=9).distinct().order_by( 'inscripcion__persona__apellido1', 'inscripcion__persona__apellido2','inscripcion__persona__nombres')
                    title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False

                    fmt = xlwt.easyxf
                    encabezado = fmt('font: height 200, bold on; border: left thin, right thin, top thin, bottom thin; align: wrap on, vert centre, horiz center;')

                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('PPL')
                    ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=admision_ppl_' + random.randint(1,10000).__str__() + '.xls'
                    columns = [
                        (u"ORDEN", 1000),
                        (u"MATRICULA", 3000),
                        (u"CEDULA", 3000),
                        (u"APELLIDOS", 6000),
                        (u"NOMBRES", 6000),
                        (u"CARRERA", 7000),
                        (u"FACULTAD", 3000),
                        (u"MODALIDAD", 3000),
                        (u"USUARIO", 3000),
                        (u"CORREO PERS.", 6000),
                        (u"CORREO INST.", 6000),
                        (u"ESTATUS", 1000),
                        (u"CRS", 9000),
                        (u"LIDER EDUCATIVO", 8000),
                        (u"CORREO DE LIDER", 6000),
                        (u"TELEFONO DE LIDER", 3000),
                        (u"FECHA DE INGRESO CRS", 3000),
                        (u"FECHA DE SALIDA CRS", 3000),
                    ]
                    row_num = 4
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], encabezado)
                        ws.col(col_num).width = columns[col_num][1]
                    row_num = 5
                    contador = 1
                    for inscripcion in inscripciones:
                        historialppl = inscripcion.persona.historialpersonappl_set.filter(status=True)
                        if len(historialppl)>0:
                            row_cont = row_num + len(historialppl) - 1
                            campo1 = str(inscripcion.persona.identificacion())
                            campo2 = str(inscripcion.persona.apellido1 + " " + inscripcion.persona.apellido2).strip()
                            campo3 = str(inscripcion.persona.nombres)
                            campo4 = str(inscripcion.coordinacion.alias)
                            campo5 = str(inscripcion.carrera.nombre_completo())
                            campo6 = str(inscripcion.modalidad.nombre)
                            campo8 = "PPL"
                            campo12 = str(inscripcion.persona.email)
                            campo13 = str(inscripcion.persona.emailinst)
                            campo14 = str(inscripcion.persona.usuario.username)
                            if MateriaAsignada.objects.filter(matricula__inscripcion=inscripcion, matriculas__gte=2).exists():
                                campo7 = "SEGUNDA"
                            else:
                                campo7 = "PRIMERA"
                            ws.write_merge(row_num, row_cont, 0, 0, contador, font_style2)
                            ws.write_merge(row_num, row_cont, 1, 1, campo7, font_style2)
                            ws.write_merge(row_num, row_cont, 2, 2, campo1, font_style2)
                            ws.write_merge(row_num, row_cont, 3, 3, campo2, font_style2)
                            ws.write_merge(row_num, row_cont, 4, 4, campo3, font_style2)
                            ws.write_merge(row_num, row_cont, 5, 5, campo5, font_style2)
                            ws.write_merge(row_num, row_cont, 6, 6, campo4, font_style2)
                            ws.write_merge(row_num, row_cont, 7, 7, campo6, font_style2)
                            ws.write_merge(row_num, row_cont, 8, 8, campo14, font_style2)
                            ws.write_merge(row_num, row_cont, 9, 9, campo12, font_style2)
                            ws.write_merge(row_num, row_cont, 10, 10, campo13, font_style2)
                            ws.write_merge(row_num, row_cont, 11, 11, campo8, font_style2)
                            for histo in historialppl:
                                ws.write(row_num, 12, histo.centrorehabilitacion, font_style2)
                                ws.write(row_num, 13, histo.lidereducativo, font_style2)
                                ws.write(row_num, 14, histo.correolidereducativo, font_style2)
                                ws.write(row_num, 15, histo.telefonolidereducativo, font_style2)
                                ws.write(row_num, 16, str(histo.fechaingreso), font_style2)
                                ws.write(row_num, 17, str(histo.fechasalida) if histo.fechasalida else '-', font_style2)
                                row_num += 1
                            contador += 1
                        else:
                            campo1 = str(inscripcion.persona.identificacion())
                            campo2 = str(inscripcion.persona.apellido1 + " " + matricula.inscripcion.persona.apellido2).strip()
                            campo3 = str(inscripcion.persona.nombres)
                            campo4 = str(inscripcion.coordinacion.alias)
                            campo5 = str(inscripcion.carrera.nombre_completo())
                            campo6 = str(inscripcion.modalidad.nombre)
                            if MateriaAsignada.objects.filter(matricula__inscripcion=inscripcion, matriculas__gte=2).exists():
                                campo7 = "SEGUNDA"
                            else:
                                campo7 = "PRIMERA"
                            campo8 = "PPL"
                            campo9 = ""
                            campo10 = ""
                            campo11 = ""
                            campo12 = str(inscripcion.persona.email)
                            campo13 = str(inscripcion.persona.emailinst)
                            campo14 = str(inscripcion.persona.usuario.username)
                            if inscripcion.persona.historialpersonappl_set.filter().exists():
                                hppl = matricula.inscripcion.persona.historialpersonappl_set.filter().order_by('-fechaingreso')[0]
                                campo9 = str(hppl.centrorehabilitacion) if hppl.centrorehabilitacion else ""
                                campo10 = str(hppl.lidereducativo) if hppl.lidereducativo else ""
                                campo11 = str(hppl.correolidereducativo) if hppl.correolidereducativo else ""
                                ws.write(row_num, 12, hppl.centrorehabilitacion, font_style2)
                                ws.write(row_num, 13, hppl.lidereducativo, font_style2)
                                ws.write(row_num, 14, hppl.correolidereducativo, font_style2)
                                ws.write(row_num, 15, hppl.telefonolidereducativo, font_style2)
                                ws.write(row_num, 16, str(hppl.fechaingreso), font_style2)
                                ws.write(row_num, 17, str(hppl.fechasalida), font_style2)

                            ws.write(row_num, 0, contador, font_style2)
                            ws.write(row_num, 1, campo7, font_style2)
                            ws.write(row_num, 2, campo1, font_style2)
                            ws.write(row_num, 3, campo2, font_style2)
                            ws.write(row_num, 4, campo3, font_style2)
                            ws.write(row_num, 5, campo5, font_style2)
                            ws.write(row_num, 6, campo4, font_style2)
                            ws.write(row_num, 7, campo6, font_style2)
                            ws.write(row_num, 8, campo14, font_style2)
                            ws.write(row_num, 9, campo12, font_style2)
                            ws.write(row_num, 10, campo13, font_style2)
                            ws.write(row_num, 11, campo8, font_style2)
                            row_num += 1
                            contador += 1
                    wb.save(response)
                    return response
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos. %s" % ex})

            elif action == 'listado_matriculado_ppl':
                try:
                    __author__ = 'Unemi'
                    matriculas = Matricula.objects.filter(status=True, nivel__periodo=periodo, inscripcion__carrera__coordinacion__id=9, inscripcion__persona__ppl=True).distinct().order_by( 'inscripcion__persona__apellido1', 'inscripcion__persona__apellido2','inscripcion__persona__nombres')
                    title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    font_style2.alignment.VERT_CENTER = True
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('PPL')
                    ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=admision_matriculados_ppl_' + random.randint(1,10000).__str__() + '.xls'
                    columns = [
                        (u"ORDEN", 1000),
                        (u"MATRICULA", 3000),
                        (u"CEDULA", 3000),
                        (u"APELLIDOS", 6000),
                        (u"NOMBRES", 6000),
                        (u"CARRERA", 7000),
                        (u"FACULTAD", 3000),
                        (u"MODALIDAD", 3000),
                        (u"USUARIO", 3000),
                        (u"CORREO PERS.", 6000),
                        (u"CORREO INST.", 6000),
                        (u"ESTATUS", 1000),
                        (u"CRS", 9000),
                        (u"LIDER EDUCATIVO", 8000),
                        (u"CORREO DE LIDER", 6000),
                        (u"TELEFONO DE LIDER", 3000),
                        (u"FECHA DE INGRESO CRS", 3000),
                        (u"FECHA DE SALIDA CRS", 3000),
                    ]
                    row_num = 4
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    row_num = 5
                    contador =1
                    for matricula in matriculas:
                        historialppl = matricula.inscripcion.persona.historialpersonappl_set.filter(status=True)
                        if historialppl.count()>0:
                            row_cont=row_num+len(historialppl)-1
                            campo1 = str(matricula.inscripcion.persona.identificacion())
                            campo2 = str(matricula.inscripcion.persona.apellido1 + " " + matricula.inscripcion.persona.apellido2).strip()
                            campo3 = str(matricula.inscripcion.persona.nombres)
                            campo4 = str(matricula.inscripcion.coordinacion.alias)
                            campo5 = str(matricula.inscripcion.carrera.nombre_completo())
                            campo6 = str(matricula.inscripcion.modalidad.nombre)
                            campo8 = "PPL"
                            campo12 = str(matricula.inscripcion.persona.email)
                            campo13 = str(matricula.inscripcion.persona.emailinst)
                            campo14 = str(matricula.inscripcion.persona.usuario.username)
                            if MateriaAsignada.objects.filter(matricula=matricula, matriculas__gte=2).exists():
                                campo7 = "SEGUNDA"
                            else:
                                campo7 = "PRIMERA"
                            ws.write_merge(row_num,row_cont, 0, 0, contador, font_style2)
                            ws.write_merge(row_num,row_cont, 1, 1, campo7, font_style2)
                            ws.write_merge(row_num,row_cont, 2, 2, campo1, font_style2)
                            ws.write_merge(row_num,row_cont, 3, 3, campo2, font_style2)
                            ws.write_merge(row_num,row_cont, 4, 4, campo3, font_style2)
                            ws.write_merge(row_num,row_cont, 5, 5, campo5, font_style2)
                            ws.write_merge(row_num,row_cont, 6, 6, campo4, font_style2)
                            ws.write_merge(row_num,row_cont, 7, 7, campo6, font_style2)
                            ws.write_merge(row_num,row_cont, 8, 8, campo14, font_style2)
                            ws.write_merge(row_num,row_cont, 9, 9, campo12, font_style2)
                            ws.write_merge(row_num,row_cont, 10, 10, campo13, font_style2)
                            ws.write_merge(row_num,row_cont, 11, 11, campo8, font_style2)
                            for histo in historialppl:
                                ws.write(row_num,12,histo.centrorehabilitacion, font_style2)
                                ws.write(row_num,13,histo.lidereducativo, font_style2)
                                ws.write(row_num,14,histo.correolidereducativo, font_style2)
                                ws.write(row_num,15,histo.telefonolidereducativo, font_style2)
                                ws.write(row_num,16,str(histo.fechaingreso), font_style2)
                                ws.write(row_num,17,str(histo.fechasalida) if histo.fechasalida else '-', font_style2)
                                row_num+=1
                            contador+=1
                        else:
                            campo1 = str(matricula.inscripcion.persona.identificacion())
                            campo2 = str(matricula.inscripcion.persona.apellido1 + " " + matricula.inscripcion.persona.apellido2).strip()
                            campo3 = str(matricula.inscripcion.persona.nombres)
                            campo4 = str(matricula.inscripcion.coordinacion.alias)
                            campo5 = str(matricula.inscripcion.carrera.nombre_completo())
                            campo6 = str(matricula.inscripcion.modalidad.nombre)
                            if MateriaAsignada.objects.filter(matricula=matricula, matriculas__gte=2).exists():
                                campo7 = "SEGUNDA"
                            else:
                                campo7 = "PRIMERA"
                            campo8 = "PPL"
                            campo9 = ""
                            campo10 = ""
                            campo11 = ""
                            campo12 = str(matricula.inscripcion.persona.email)
                            campo13 = str(matricula.inscripcion.persona.emailinst)
                            campo14 = str(matricula.inscripcion.persona.usuario.username)
                            if matricula.inscripcion.persona.historialpersonappl_set.filter().exists():
                                hppl = matricula.inscripcion.persona.historialpersonappl_set.filter().order_by('-fechaingreso')[0]
                                campo9 = str(hppl.centrorehabilitacion) if hppl.centrorehabilitacion else ""
                                campo10 = str(hppl.lidereducativo) if hppl.lidereducativo else ""
                                campo11 = str(hppl.correolidereducativo) if hppl.correolidereducativo else ""
                                ws.write(row_num, 12, hppl.centrorehabilitacion, font_style2)
                                ws.write(row_num, 13, hppl.lidereducativo, font_style2)
                                ws.write(row_num, 14, hppl.correolidereducativo, font_style2)
                                ws.write(row_num, 15, hppl.telefonolidereducativo, font_style2)
                                ws.write(row_num, 16, str(hppl.fechaingreso), font_style2)
                                ws.write(row_num, 17, str(hppl.fechasalida), font_style2)

                            ws.write(row_num, 0, contador, font_style2)
                            ws.write(row_num, 1, campo7, font_style2)
                            ws.write(row_num, 2, campo1, font_style2)
                            ws.write(row_num, 3, campo2, font_style2)
                            ws.write(row_num, 4, campo3, font_style2)
                            ws.write(row_num, 5, campo5, font_style2)
                            ws.write(row_num, 6, campo4, font_style2)
                            ws.write(row_num, 7, campo6, font_style2)
                            ws.write(row_num, 8, campo14, font_style2)
                            ws.write(row_num, 9, campo12, font_style2)
                            ws.write(row_num, 10, campo13, font_style2)
                            ws.write(row_num, 11, campo8, font_style2)
                            row_num += 1
                            contador += 1
                    wb.save(response)
                    return response
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos. %s" % ex})

            elif action == 'notasmoodle':
                try:
                    data['title'] = u'Listado de Asignaturas'
                    url_vars = ''
                    filtro = Q(status=True, nivel__periodo=periodo, idcursomoodle__gt=0, asignaturamalla__malla__carrera__coordinacion__id=9,materiaasignada__status=True)
                    idcarrera = Materia.objects.filter(filtro).distinct().values_list('asignaturamalla__malla__carrera__id', flat=True)
                    carreras = Carrera.objects.filter(id__in=idcarrera, status=True)
                    if 'estado' in request.GET:
                        data['estado'] = estado = request.GET['estado']
                        filtro = filtro & Q(cerrado=(estado == '2'))
                        url_vars += "&estado={}".format(estado)
                    if 's' in request.GET:
                        data['search'] = search = request.GET['s'].strip()
                        filtro = filtro & (Q(asignatura__nombre__icontains=search) | Q(paralelo__icontains=search))
                        url_vars += "&s={}".format(search)
                    if 'car' in request.GET:
                        data['car'] = carrera = int(request.GET['car'])
                        filtro = filtro & Q(asignaturamalla__malla__carrera__id=carrera)
                        url_vars += "&car={}".format(carrera)
                    materias = Materia.objects.filter(filtro).distinct().order_by('asignatura', '-cerrado')

                    paging = MiPaginador(materias, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                            url_vars += "&page={}".format(p)
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['materias'] = page.object_list
                    data['total'] = materias.count()
                    data['url_vars'] = url_vars
                    data['carreraslist'] = carreras
                    return render(request, "inscripciones_admision/notasmoodle.html", data)
                except Exception as e:
                    print(e)

            elif action == 'getnotasmoodle':
                try:
                    data['title'] = u'Notas de moodle'
                    lista = []
                    data['materia'] = materia= Materia.objects.get(pk=int(encrypt(request.GET['id'])))
                    data['inscritos'] = inscritos=materia.asignados_a_esta_materia_moodle()
                    data['utiliza_validacion_calificaciones'] = variable_valor('UTILIZA_VALIDACION_CALIFICACIONES')
                    # data['habilitado_ingreso_calificaciones'] = profesor.habilitado_ingreso_calificaciones()
                    data['habilitado_ingreso_calificaciones'] = True
                    url_vars = ''
                    if 'estado' in request.GET:
                        url_vars += "&estado={}".format(request.GET['estado'])
                    if 's' in request.GET:
                        url_vars += "&s={}".format(request.GET['s'])
                    if 'page' in request.GET:
                        url_vars += "&page={}".format(request.GET['page'])
                    data['url_vars'] = url_vars
                    return render(request, "inscripciones_admision/getnotasmoodle.html", data)
                except Exception as ex:
                    return HttpResponseRedirect("/inscripciones_admision?action=notasmoodle")

            elif action == 'notasmateria':
                data['title'] = u'Calificaciones de Estudiantes'
                # if persona.id not in (42552, 41508, 38029, 42551, 42553, 54, 3545, 4379):
                #     if periodo.tipo.id != 3:
                #         if str(persona.id) not in variable_valor('CALIFICAR_FUERA_UNEMI'):
                #             # if not inhouse_check(request) and VALIDATE_IPS:
                #             if not inhouse_check(request) and variable_valor('VALIDATE_IPS'):
                #                 log(u'Bloqueo de ip externa por ingreso de notas: %s' % get_client_ip(request), request, "add")
                #                 return HttpResponseRedirect("/?info=No puede ingresar las calificaciones fuera de la institucion.")

                hoy = datetime.now().date()
                data['id'] = id = int(encrypt(request.GET['id']))
                materia = Materia.objects.get(status=True, id=id)
                otrasmaterias = []
                if periodo.ocultarmateria:
                    materia = False
                data['materia'] = materia

                data['utiliza_validacion_calificaciones'] = variable_valor('UTILIZA_VALIDACION_CALIFICACIONES')
                # data['habilitado_ingreso_calificaciones'] = profesor.habilitado_ingreso_calificaciones()
                data['habilitado_ingreso_calificaciones'] = True
                # data['profesor'] = profesor
                url_vars = ''
                if 'estado' in request.GET:
                    url_vars += "&estado={}".format(request.GET['estado'])
                if 's' in request.GET:
                    url_vars += "&s={}".format(request.GET['s'])
                if 'page' in request.GET:
                    url_vars += "&page={}".format(request.GET['page'])
                data['url_vars'] = url_vars
                try:
                    return render(request, "inscripciones_admision/notasmateria.html", data)
                except Exception as ex:
                    pass

            elif action == 'segmento':
                    try:
                        data['materia'] = materia = Materia.objects.get(pk=int(encrypt(request.GET['id'])))
                        data['cronograma'] = materia.cronogramacalificaciones()
                        data['usacronograma'] = materia.usaperiodocalificaciones
                        # data['usa_evaluacion_integral'] = USA_EVALUACION_INTEGRAL
                        # data['validardeuda'] = PAGO_ESTRICTO
                        # data['incluyedatos'] = DATOS_ESTRICTO
                        data['dentro_fechas'] = materia.fin >= datetime.now().date()
                        data['auditor'] = False
                        if materia.asignaturamalla.malla.carrera.coordinacion_carrera().id == 9:
                            data['reporte_0'] = obtener_reporte('acta_calificaciones_admision')  # obtener_reporte('acta_notas_admision')
                        else:
                            data['reporte_0'] = obtener_reporte('acta_notas')
                        data['reporte_1'] = obtener_reporte('lista_control_calificaciones')
                        data['reporte_2'] = obtener_reporte('acta_notas_parcial')
                        bandera = False
                        if PlanificacionMateria.objects.filter(materia=materia, paraevaluacion=True).exists():
                            bandera = True
                        data['bandera'] = bandera
                        return render(request, "inscripciones_admision/segmentonotasmoddle.html", data)
                    except Exception as ex:
                        pass

            elif action == 'reporte_acta_calificaciones':
                try:
                    materia_id = int(encrypt(request.GET['materia']))
                    # reporte_id = request.GET['reporte']
                    reporte0 = obtener_reporte('acta_calificaciones_admision')
                    reporte = Reporte.objects.get(pk=reporte0.pk)
                    base_url = request.META['HTTP_HOST']
                    d = datetime.now()
                    pdfname = reporte.nombre + d.strftime('%Y%m%d_%H%M%S')
                    tipo = 'pdf'
                    paRequest = {
                                    'materia': str(materia_id),
                                    'imp_logo':True,
                                    'imp_encabezado':True,
                                    'imp_fecha':True,
                                    'imp_membretada':False,
                                    'url_qr':unicode(base_url + "/".join([MEDIA_URL, 'documentos', 'userreports', elimina_tildes(request.user.username), pdfname + "." + tipo]))
                                }
                    d = run_report_v1(reporte=reporte,tipo=tipo , paRequest=paRequest, request=request)
                    if not d['isSuccess']:
                        raise NameError(d['mensaje'])
                    else:
                        return ok_json({"r": d['mensaje'], 'reportfile':d['data']['reportfile']})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return bad_json(mensaje="Error, al generar el reporte. %s" % ex.__str__())

            elif action == 'reporte_acta_calificacionesnotas':
                try:
                    materia_id = request.GET['materia']
                    reporte_id = request.GET['reporte']
                    reporte = Reporte.objects.get(pk=reporte_id)
                    base_url = request.META['HTTP_HOST']
                    d = datetime.now()
                    pdfname = reporte.nombre + d.strftime('%Y%m%d_%H%M%S')
                    tipo = 'pdf'
                    paRequest = {
                                    'materia': materia_id,
                                    'imp_logo':True,
                                    'imp_encabezado':True,
                                    'imp_fecha':True,
                                    'imp_membretada':False,
                                    'url_qr':unicode(base_url + "/".join([MEDIA_URL, 'documentos', 'userreports', elimina_tildes(request.user.username), pdfname + "." + tipo]))
                                }
                    d = run_report_v1(reporte=reporte,tipo=tipo , paRequest=paRequest, request=request)
                    if not d['isSuccess']:
                        raise NameError(d['mensaje'])
                    else:
                        return ok_json({"r": d['mensaje'], 'reportfile':d['data']['reportfile']})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return bad_json(mensaje="Error, al generar el reporte. %s" % ex.__str__())

            elif action == 'sedeexamenes':
                try:
                    from inno.models import MatriculaSedeExamen
                    data['title'] = u'Sedes del estudiante de exámenes'
                    data['matricula'] = matricula = Matricula.objects.get(pk=request.GET['id'])
                    data['inscripcion'] = matricula.inscripcion
                    data['ret'] = request.GET['ret']
                    data['eMatriculaSedeExamenes'] = MatriculaSedeExamen.objects.filter(status=True, matricula=matricula)
                    return render(request, "inscripciones_admision/sedesexamenes/view.html", data)
                except Exception as ex:
                    pass

            elif action == 'addsedeexamen':
                try:
                    from inno.models import MatriculaSedeExamen
                    data['title'] = u"Adicionar sede de examen"
                    data['matricula'] = matricula = Matricula.objects.get(pk=request.GET['id'])
                    data['inscripcion'] = matricula.inscripcion
                    data['ret'] = request.GET['ret']
                    # data['sedes'] = Sede.objects.filter(status=True)
                    form = MatriculaSedeExamenForm()
                    mismaterias = matricula.mismaterias()
                    modeloevaluativos = mismaterias.values_list('materia__modeloevaluativo_id', flat=True) if mismaterias.values("id").exists() else []
                    form.examen(modeloevaluativos)
                    data['form'] = form
                    return render(request, "inscripciones_admision/sedesexamenes/new.html", data)
                except Exception as ex:
                    pass

            elif action == 'editsedeexamen':
                try:
                    from inno.models import MatriculaSedeExamen
                    data['title'] = u"Editar sede de examen"
                    data['eMatriculaSedeExamen'] = eMatriculaSedeExamen = MatriculaSedeExamen.objects.get(pk=request.GET['id'])
                    data['matricula'] = matricula = eMatriculaSedeExamen.matricula
                    data['inscripcion'] = matricula.inscripcion
                    data['ret'] = request.GET['ret']
                    form = MatriculaSedeExamenForm(initial={'sede': eMatriculaSedeExamen.sede,
                                                            'detallemodeloevaluativo': eMatriculaSedeExamen.detallemodeloevaluativo
                                                            }
                                                   )
                    mismaterias = matricula.mismaterias()
                    modeloevaluativos = mismaterias.values_list('materia__modeloevaluativo_id', flat=True) if mismaterias.values("id").exists() else []
                    form.examen(modeloevaluativos)
                    data['form'] = form
                    return render(request, "inscripciones_admision/sedesexamenes/edit.html", data)
                except Exception as ex:
                    pass


            return HttpResponseRedirect(request.path)
        else:
            try:
                data['title'] = u'Listado de inscripciones Admisión'
                search = None
                ids = None
                carreraselect = 0
                modalidadselect = 0
                inscripciones = Inscripcion.objects.filter(carrera__coordinacion__id=9).order_by('persona__apellido1', 'persona__apellido2')
                if 'id' in request.GET:
                    ids = request.GET['id']
                    inscripciones = inscripciones.filter(id=ids)
                if 's' in request.GET:
                    search = request.GET['s'].strip()
                    ss = search.split(' ')
                    if len(ss) == 1:
                        inscripciones = inscripciones.filter(Q(persona__nombres__icontains=search) |
                                                             Q(persona__apellido1__icontains=search) |
                                                             Q(persona__apellido2__icontains=search) |
                                                             Q(persona__cedula__icontains=search) |
                                                             Q(persona__pasaporte__icontains=search) |
                                                             Q(identificador__icontains=search) |
                                                             Q(inscripciongrupo__grupo__nombre__icontains=search) |
                                                             Q(persona__usuario__username__icontains=search)).distinct()
                    else:
                        inscripciones = inscripciones.filter(Q(persona__apellido1__icontains=ss[0]) &
                                                             Q(persona__apellido2__icontains=ss[1])).distinct()
                if 'c' in request.GET:
                    carreraselect = int(request.GET['c'])
                    if carreraselect > 0:
                        inscripciones = inscripciones.filter(carrera_id=carreraselect)

                if 'm' in request.GET:
                    modalidadselect = int(request.GET['m'])
                    if modalidadselect > 0:
                        inscripciones = inscripciones.filter(modalidad_id=modalidadselect)

                ppl = 0
                if 'ppl' in request.GET and int(request.GET['ppl']) > 0:
                    ppl = int(request.GET['ppl'])
                    inscripciones = inscripciones.filter(persona__ppl=int(request.GET['ppl']) == 1)

                paging = MiPaginador(inscripciones, 15)
                p = 1
                try:
                    paginasesion = 1
                    if 'paginador' in request.session:
                        paginasesion = int(request.session['paginador'])
                    if 'page' in request.GET:
                        p = int(request.GET['page'])
                    else:
                        p = paginasesion
                    try:
                        page = paging.page(p)
                    except:
                        p = 1
                    page = paging.page(p)
                except:
                    page = paging.page(p)
                request.session['paginador'] = p
                data['paging'] = paging
                data['rangospaging'] = paging.rangos_paginado(p)
                data['page'] = page
                data['search'] = search if search else ""
                data['ids'] = ids if ids else ""
                data['ppl'] = ppl
                data['inscripciones'] = page.object_list
                data['utiliza_grupos_alumnos'] = UTILIZA_GRUPOS_ALUMNOS
                data['clave'] = DEFAULT_PASSWORD
                data['reporte_0'] = obtener_reporte('listado_vinculacion_estudiante')
                data['reporte_1'] = obtener_reporte('ficha_inscripcion')
                data['reporte_2'] = obtener_reporte('certificado_inscripcion')
                data['reporte_3'] = obtener_reporte('certificado_matricula_alumno_plus')
                data['reporte_4'] = obtener_reporte('hoja_vida')
                data['reporte_5'] = obtener_reporte('certificado_matricula_alumno')
                data['reporte_6'] = obtener_reporte("record_alumno")
                data['reporte_7'] = obtener_reporte("registro_matriculacion")
                data['reporte_8'] = obtener_reporte("compromiso_pago")
                data['control_unico_credenciales'] = CONTROL_UNICO_CREDENCIALES
                data['usa_tipos_inscripciones'] = USA_TIPOS_INSCRIPCIONES
                data['matriculacion_libre'] = MATRICULACION_LIBRE
                data['periodo'] = request.session['periodo']
                # data['carreras'] = carreras = Carrera.objects.filter(id__in=Matricula.objects.values_list('inscripcion__carrera__id', flat=True).filter(nivel__periodo=periodo, inscripcion__carrera__coordinacion__id=9)).order_by('modalidad')
                data['carreras'] = carreras = Carrera.objects.filter(status=True, coordinacion__id=9).order_by('modalidad')
                data['carreraselect'] = carreraselect
                data['modalidadselect'] = modalidadselect
                data['cant'] = inscripciones.count()
                # data['pais'] = Pais.objects.filter(
                #     id__in=inscripciones.values_list('persona__pais__id', flat=True).filter(
                #         persona__pais__status=True))
                data['totalemailenviados'] = inscripciones.filter(modalidad_id=3, envioemail=True).count()
                data['totalonline'] = inscripciones.filter(modalidad_id=3, confimacion_online=True).count()
                return render(request, "inscripciones_admision/view.html", data)
            except Exception as ex:
                pass
