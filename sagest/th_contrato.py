# -*- coding: UTF-8 -*-
import json
import os
import io

import requests
import xlsxwriter
import xlwt
from django.db.models import Q
from django.forms.models import model_to_dict
from django.template.loader import get_template
from openpyxl import load_workbook

from plan.forms import PlanAccionPersonaPlanThForm
from postulate.models import PersonaAplicarPartida, PersonaPeriodoConvocatoria
from sga.adm_convenioempresa import buscar_dicc
from sga.funciones import MiPaginador, log, generar_nombre

from datetime import datetime

import xlsxwriter
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.http import HttpResponseRedirect, JsonResponse,HttpResponse
from django.shortcuts import render
from decorators import secure_module
from sagest.forms import CampoContratoForm, ContratosForm, ArchivoContratoForm, EditContratosForm, \
    MantenimientoNombreForm, ArchivoImportarContratoForm
from sagest.models import CamposContratos, Contratos, ContratosCamposSeleccion, TipoContrato, ArchivoContrato, \
    Departamento, DenominacionPuesto, OtroRegimenLaboral, PersonaContratos
from sagest.models import ContratoPersona, ContratoPersonaDetalle
from settings import SITE_ROOT
from sga.commonviews import adduserdata
from sga.funciones import log
from docx import Document
from docx.shared import Pt

from sga.models import Persona


def rango_anios():
    if Contratos.objects.exists():
        inicio = datetime.now().year
        fin = Contratos.objects.order_by('anio')[0].anio
        return range(inicio, fin , -1)
    return [datetime.now().date().year]


def fecha_letra(valor):
    mes = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    d = int(valor[0:2])
    m = int(valor[3:5])
    a = int(valor[6:10])
    if d == 1:
        return u"al %s día del mes de %s del %s" % (numero_a_letras(d), str(mes[m - 1]), numero_a_letras(a))
    else:
        return u"a los %s días del mes de %s del %s" % (numero_a_letras(d), str(mes[m - 1]), numero_a_letras(a))

MONEDA_SINGULAR = 'dolar'
MONEDA_PLURAL = 'dolares'
CENTIMOS_SINGULAR = 'centavo'
CENTIMOS_PLURAL = 'centavos'
MAX_NUMERO = 999999999999

UNIDADES = (
    'cero',
    'uno',
    'dos',
    'tres',
    'cuatro',
    'cinco',
    'seis',
    'siete',
    'ocho',
    'nueve'
)

DECENAS = (
    'diez',
    'once',
    'doce',
    'trece',
    'catorce',
    'quince',
    'dieciseis',
    'diecisiete',
    'dieciocho',
    'diecinueve'
)

DIEZ_DIEZ = (
    'cero',
    'diez',
    'veinte',
    'treinta',
    'cuarenta',
    'cincuenta',
    'sesenta',
    'setenta',
    'ochenta',
    'noventa'
)

CIENTOS = (
    '_',
    'ciento',
    'doscientos',
    'trescientos',
    'cuatroscientos',
    'quinientos',
    'seiscientos',
    'setecientos',
    'ochocientos',
    'novecientos'
)


def numero_a_letras(numero):
    numero_entero = int(numero)
    if numero_entero > MAX_NUMERO:
         raise OverflowError('Número demasiado alto')
    if numero_entero < 0:
        return 'menos %s' % numero_a_letras(abs(numero))
    letras_decimal = ''
    parte_decimal = int(round((abs(numero) - abs(numero_entero)) * 100))
    if parte_decimal > 9:
        letras_decimal = 'punto %s' % numero_a_letras(parte_decimal)
    elif parte_decimal > 0:
        letras_decimal = 'punto cero %s' % numero_a_letras(parte_decimal)
    if numero_entero <= 99:
        resultado = leer_decenas(numero_entero)
    elif numero_entero <= 999:
        resultado = leer_centenas(numero_entero)
    elif numero_entero <= 999999:
        resultado = leer_miles(numero_entero)
    elif numero_entero <= 999999999:
        resultado = leer_millones(numero_entero)
    else:
        resultado = leer_millardos(numero_entero)
    resultado = resultado.replace('uno mil', 'un mil')
    resultado = resultado.strip()
    resultado = resultado.replace(' _ ', ' ')
    resultado = resultado.replace('  ', ' ')
    if parte_decimal > 0:
        resultado = '%s %s' % (resultado, letras_decimal)
    return resultado


def numero_a_moneda(numero):
    if '.' in numero:
        posicion = numero.split('.')
        numero_entero = int(posicion[0])
        parte_decimal = 0
        if posicion[1] != '':
            parte_decimal = int(posicion[1])
    else:
        numero_entero = int(numero)
        parte_decimal = 0
    centimos = ''
    if parte_decimal == 1:
        centimos = CENTIMOS_SINGULAR
    else:
        centimos = CENTIMOS_PLURAL
    moneda = ''
    if numero_entero == 1:
        moneda = MONEDA_SINGULAR
    else:
        moneda = MONEDA_PLURAL
    letras = numero_a_letras(numero_entero)
    letras = letras.replace('uno', 'un')
    letras_decimal = u'%s/100 DOLARES DE LOS ESTADOS UNIDOS DE NORTE AMÉRICA' % (str(parte_decimal))
    letras = u'%s %s' % (letras, letras_decimal)
    return letras


def leer_decenas(numero):
    if numero < 10:
        return UNIDADES[numero]
    decena, unidad = divmod(numero, 10)
    if unidad == 0:
        resultado = DIEZ_DIEZ[decena]
    else:
        if numero <= 19:
            resultado = DECENAS[unidad]
        elif numero <= 29:
            resultado = 'veinti%s' % UNIDADES[unidad]
        else:
            resultado = DIEZ_DIEZ[decena]
            if unidad > 0:
                resultado = '%s y %s' % (resultado, UNIDADES[unidad])
    return resultado


def leer_centenas(numero):
    centena, decena = divmod(numero, 100)
    if numero == 0:
        resultado = 'cien'
    else:
        resultado = CIENTOS[centena]
        if decena > 0:
            resultado = '%s %s' % (resultado, leer_decenas(decena))
    return resultado


def leer_miles(numero):
    millar, centena = divmod(numero, 1000)
    resultado = ''
    if millar == 1:
        resultado = ''
    if 2 <= millar <= 9:
        resultado = UNIDADES[millar]
    elif 10 <= millar <= 99:
        resultado = leer_decenas(millar)
    elif 100 <= millar <= 999:
        resultado = leer_centenas(millar)
    resultado = '%s mil' % resultado
    if centena > 0:
        resultado = '%s %s' % (resultado, leer_centenas(centena))
    return resultado


def leer_millones(numero):
    millon, millar = divmod(numero, 1000000)
    resultado = ''
    if millon == 1:
        resultado = ' un millon '
    if 2 <= millon <= 9:
        resultado = UNIDADES[millon]
    elif 10 <= millon <= 99:
        resultado = leer_decenas(millon)
    elif 100 <= millon <= 999:
        resultado = leer_centenas(millon)
    if millon > 1:
        resultado = '%s millones' % resultado
    if (millar > 0) and (millar <= 999):
        resultado = '%s %s' % (resultado, leer_centenas(millar))
    elif (millar >= 1000) and (millar <= 999999):
        resultado = '%s %s' % (resultado, leer_miles(millar))
    return resultado


def leer_millardos(numero):
    millardo, millon = divmod(numero, 1000000)
    return '%s millones %s' % (leer_miles(millardo), leer_millones(millon))


@login_required(redirect_field_name='ret', login_url='/loginsagest')
@secure_module
@transaction.atomic()
def view(request):
    data = {}
    anios = rango_anios()
    if 'aniotipocontrato' in request.session:
        data['aniotipocontrato'] = request.session['aniotipocontrato']
    else:
        request.session['aniotipocontrato'] = anios[0]
    data['anioselect'] = anioselect = request.session['aniotipocontrato']
    adduserdata(request, data)
    persona = request.session['persona']
    usuario = request.user
    if request.method == 'POST':
        action = request.POST['action']

        if action == 'addcampos':
            try:
                form = CampoContratoForm(request.POST, request.FILES)
                if form.is_valid():
                    if CamposContratos.objects.filter(identificador=(form.cleaned_data['identificador']).upper(),status=True).exists():
                        return JsonResponse({"result": "bad", "mensaje": "El identificador ya existe."})

                    registro = CamposContratos(descripcion=form.cleaned_data['descripcion'],
                                               tipo=form.cleaned_data['tipo'],
                                               script=form.cleaned_data['script'],
                                               identificador=form.cleaned_data['identificador'],
                                               fijo=form.cleaned_data['fijo'])
                    registro.save(request)
                    log(u'Registro nuevo de campos contrato: %s' % registro, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                     raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": "Error al guardar los datos, dato duplicado."})

        if action == 'addcontrato':
            try:
                form = ContratosForm(request.POST, request.FILES)
                if form.is_valid():
                    if not json.loads(request.POST['lista_items1']):
                        raise NameError('Ningún postulante agregado')
                    for postulante in json.loads(request.POST['lista_items1']):
                        id=postulante['id']
                        if not request.POST['departamento_'+id]:
                            raise NameError('Existe uno o varios registros sin seleccionar un departamento.')
                        if not request.POST['rmu_'+id]:
                            raise NameError('Existe uno o varios registros sin describir su RMU.')

                        p_periodo = PersonaPeriodoConvocatoria.objects.get(pk=id)
                        departamento=Departamento.objects.get(pk=request.POST['departamento_'+id])
                        contrato = PersonaContratos(
                            persona=p_periodo.persona,
                            estado=True,
                            explicacion=form.cleaned_data['explicacion'],
                            fechainicio=form.cleaned_data['fechainicio'],
                            fechafin=form.cleaned_data['fechafin'],
                            relacionies=form.cleaned_data['relacionies'],
                            regimenlaboral=form.cleaned_data['regimenlaboral'],
                            remuneracion = request.POST['rmu_'+id],
                            unidad = str(departamento),
                            unidadorganica = departamento,
                            denominacionpuesto = p_periodo.denominacionpuesto,
                            cargo = str(p_periodo.denominacionpuesto.descripcion) if p_periodo.denominacionpuesto else '' ,
                            estadocontrato=1,
                            periodopostulate=p_periodo

                        )
                        contrato.save(request)
                        log(u'Registro nuevo contrato a: %s' % p_periodo, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                         "mensaje": "Error en el formulario"})

            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": str(ex)})


        if action == 'editcontrato':
            try:
                form = EditContratosForm(request.POST, request.FILES)
                contrato = PersonaContratos.objects.get(pk=request.POST['id'])
                if form.is_valid():

                    contrato.explicacion=form.cleaned_data['explicacion']
                    contrato.fechainicio=form.cleaned_data['fechainicio']
                    contrato.fechafin=form.cleaned_data['fechafin']
                    contrato.numerodocumento = form.cleaned_data['numerodocumento']
                    contrato.denominacionpuesto = form.cleaned_data['denominacionpuesto']
                    contrato.relacionies=form.cleaned_data['relacionies']
                    contrato.regimenlaboral=form.cleaned_data['regimenlaboral']
                    contrato.remuneracion=form.cleaned_data['remuneracion']
                    contrato.unidadorganica=form.cleaned_data['unidadorganica']
                    contrato.estado=True
                    if form.cleaned_data['unidadorganica']:
                        contrato.unidad = str(contrato.unidadorganica)
                    if form.cleaned_data['denominacionpuesto']:
                        contrato.cargo = str(contrato.denominacionpuesto)
                    contrato.save(request)

                    log(u'Editó contrato : %s' % contrato, request, "edit")
                    return JsonResponse({"result": False}, safe=False)
                else:
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                         "message": "Error en el formulario"})

            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)




        elif action == 'deletecontratopersona':
            try:
                registro = ContratoPersona.objects.get(pk=request.POST['id'], status=True)
                registro.status = False
                registro.save(request)
                log(u'Elimino contrato persona: %s' % registro, request, "del")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": "Error al eliminar los datos."})




        return JsonResponse({"result": "bad", "mensaje": "Solicitud Incorrecta."})
    else:
        if 'action' in request.GET:
            data['action'] = action = request.GET['action']

            if action == 'addcontrato':
                try:
                    data['title'] = u'Adicionar Contrato'
                    data['form'] = ContratosForm()
                    # contratos = PersonaContratos.objects.filter(partidapostulate__isnull=False,status=True).values_list('partidapostulate_id', flat=True )
                    # ganadores = PersonaAplicarPartida.objects.filter(status=True, esganador=True).exclude(id__in=contratos)
                    contratos = PersonaContratos.objects.filter(periodopostulate__isnull=False,status=True).values_list('periodopostulate_id', flat=True )
                    p_contratacion = PersonaPeriodoConvocatoria.objects.filter(status=True, estado=2).exclude(id__in=contratos)
                    data['ganadores'] = p_contratacion
                    return render(request, "th_contrato/addcontrato.html", data)
                except Exception as ex:
                    pass

            if action == 'detalles':
                try:
                    data['contrato'] = contrato = PersonaContratos.objects.get(pk=request.GET['id'])
                    template = get_template("th_contrato/modal/detalles.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass


            if action == 'buscarpersonas':
                try:
                    id = request.GET['id']
                    q = request.GET['q'].upper().strip()
                    s = q.split(" ")
                    querybase = Persona.objects.filter(administrativo__isnull=False)
                    if len(s) == 1:
                        per = querybase.filter((Q(nombres__icontains=q) | Q(apellido1__icontains=q) | Q(
                            cedula__icontains=q) | Q(apellido2__icontains=q) | Q(cedula__contains=q)),
                                               Q(status=True)).distinct()[:15]
                    elif len(s) == 2:
                        per = querybase.filter((Q(apellido1__contains=s[0]) & Q(apellido2__contains=s[1])) |
                                               (Q(nombres__icontains=s[0]) & Q(nombres__icontains=s[1])) |
                                               (Q(nombres__icontains=s[0]) & Q(apellido1__contains=s[1]))).filter(
                            status=True).distinct()[:15]
                    else:
                        per = querybase.filter(
                            (Q(nombres__contains=s[0]) & Q(apellido1__contains=s[1]) & Q(apellido2__contains=s[2])) |
                            (Q(nombres__contains=s[0]) & Q(nombres__contains=s[1]) & Q(
                                apellido1__contains=s[2]))).filter(status=True).distinct()[:15]
                    data = {"result": "ok",
                            "results": [{"id": x.id, "name": "{} - {}".format(x.cedula, x.nombre_completo())} for x in
                                        per]}
                    return JsonResponse(data)
                except Exception as ex:
                    pass


            if action == 'editcontrato':
                try:
                    data['title'] = u'Modificación de Contrato'
                    data['filtro'] = contrato = PersonaContratos.objects.get(pk=request.GET['id'])
                    data['form'] = form = EditContratosForm(initial=model_to_dict(contrato))
                    template = get_template("th_contrato/modal/formcontrato.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'configuraciones':
                try:
                    data['title'] = u'Configuraciones'
                    data['campos'] = CamposContratos.objects.filter(status=True).order_by('-id')
                    data['tipos'] = TipoContrato.objects.filter(status=True).order_by('-id')
                    data['modeloscontratos'] = Contratos.objects.filter(status=True).order_by('-anio','-pk')
                    return render(request, "th_contrato/configurar.html", data)
                except Exception as ex:
                    pass


            if action == 'deletecampo':
                try:
                    data['title'] = u'Eliminar Campo'
                    data['campo'] = CamposContratos.objects.get(pk=int(request.GET['id']))
                    return render(request, "th_contrato/deletecampo.html", data)
                except:
                    pass

            if action == 'deletecontrato':
                try:
                    data['title'] = u'Eliminar Contrato Plantilla'
                    data['contrato'] = Contratos.objects.get(pk=int(request.GET['id']))
                    return render(request, "th_contrato/deletecontrato.html", data)
                except:
                    pass

            if action == 'deletecontratopersona':
                try:
                    data['title'] = u'Eliminar Contrato Persona'
                    data['contrato'] = ContratoPersona.objects.get(pk=int(request.GET['id']))
                    return render(request, "th_contrato/deletecontratopersona.html", data)
                except:
                    pass

            if action == 'addarchivocontrato':
                try:
                    contrato =Contratos.objects.get(pk=int(request.GET['id']))
                    if contrato.archivo:
                        data['filtro'] = filtro = Contratos.objects.get(pk=int(request.GET['id']))
                        data['idcontratos'] = request.GET['id']
                        data['form2'] = ArchivoContratoForm(initial=model_to_dict(filtro))
                    else:
                        form2 = ArchivoContratoForm()
                        data['idcontratos'] = request.GET['id']
                        data['form2'] = form2
                    # return render(request, "th_contrato/addarchivocontrato.html", data)
                    template = get_template("th_contrato/addarchivocontrato.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except:
                    pass
            if action == 'addtipo':
                try:
                    data['form2'] = MantenimientoNombreForm()
                    template = get_template("th_contrato/modal/tipo.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'edittipo':
                try:
                    data['id'] = request.GET['id']
                    data['filtro'] = filtro = TipoContrato.objects.get(pk=request.GET['id'])
                    data['form2'] = MantenimientoNombreForm(initial=model_to_dict(filtro))
                    template = get_template("th_contrato/modal/tipo.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'deletetipo':
                try:
                    data['title'] = u'Eliminar Tipo de Contrato'
                    data['tipo'] = TipoContrato.objects.get(pk=int(request.GET['id']))
                    return render(request, "th_contrato/deletetipo.html", data)
                except:
                    pass

            if action == 'subir':
                try:
                    data['form2'] = ArchivoImportarContratoForm()
                    template = get_template("th_contrato/modal/archivoimportar.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'descargaplantilla':
                try:
                    contrato = Contratos.objects.get(pk=request.GET['id'])
                    campos = contrato.contratoscamposseleccion_set.filter(status=True)
                    __author__ = 'Unemi'

                    output = io.BytesIO()
                    workbook = xlsxwriter.Workbook(output)
                    ws = workbook.add_worksheet('plantilla_')
                    ws.set_column(0, 100, 60)

                    formatoceldagris = workbook.add_format(
                        {'align': 'center', 'border': 1, 'text_wrap': True, 'fg_color': '#B6BFC0'})

                    ws.write(0, 0, 'CEDULA', formatoceldagris)
                    ws.write(1, 0, 'CEDULA', formatoceldagris)
                    ws.write(0, 1, 'UNIDAD ORGÁNICA', formatoceldagris)
                    ws.write(1, 1, 'DEPARTAMENTO', formatoceldagris)
                    ws.write(0, 2, 'DENOMINACION PUESTO', formatoceldagris)
                    ws.write(1, 2, 'PUESTO', formatoceldagris)
                    ws.write(0, 3, 'REMUNERACIÓN', formatoceldagris)
                    ws.write(1, 3, 'RMU', formatoceldagris)
                    cont = 4
                    for campo in campos:
                        ws.write(0, cont, str(campo.campos), formatoceldagris)
                        ws.write(1, cont,
                                 str(campo.campos.identificador if campo.campos.identificador else 'INDEFINIDO'),
                                 formatoceldagris)
                        cont += 1
                    workbook.close()
                    output.seek(0)
                    filename = 'plantilla_%s.xlsx' % (contrato.descripcion)
                    response = HttpResponse(output,
                                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = 'attachment; filename=%s' % filename
                    return response
                except Exception as ex:
                    pass

            elif action == 'importaciones':
                try:
                    data['title'] = u'Importaciones'
                    search = None
                    ids = None
                    if 's' in request.GET:
                        search = request.GET['s']
                        importacion = ArchivoContrato.objects.filter(Q(contrato__descripcion__icontains=search) |
                                                                     Q(contrato__regimenlaboral__nombre__icontains=search)).order_by(
                            '-id')
                    elif 'id' in request.GET:
                        ids = request.GET['id']
                        importacion = ArchivoContrato.objects.filter(id=ids).order_by('-id')
                    else:
                        importacion = ArchivoContrato.objects.all().order_by('-id')
                    paging = MiPaginador(importacion, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['importaciones'] = page.object_list
                    return render(request, "th_contrato/importaciones.html", data)
                except Exception as ex:
                    pass

            elif action == 'generarreporte':
                try:
                    filtro = Q(fechainicio__range=(request.GET['desde'], request.GET['hasta']))
                    contrato = Contratos.objects.get(pk=request.GET['modelo'])
                    campos = contrato.contratoscamposseleccion_set.filter(status=True)
                    contratopersona = ContratoPersona.objects.filter(filtro, status=True, contrato=contrato)

                    __author__ = 'Unemi'

                    output = io.BytesIO()
                    workbook = xlsxwriter.Workbook(output)
                    ws = workbook.add_worksheet('plantilla_')
                    ws.set_column(0, 100, 60)

                    formatoceldagris = workbook.add_format(
                        {'align': 'center', 'border': 1, 'text_wrap': True, 'fg_color': '#B6BFC0'})
                    formatoceldaleft = workbook.add_format({'text_wrap': True, 'align': 'left'})

                    ws.write(0, 0, 'CODIGO CONTRATO', formatoceldagris)
                    ws.write(0, 1, 'FECHA CONTRATO', formatoceldagris)
                    ws.write(0, 2, 'CEDULA', formatoceldagris)
                    ws.write(0, 3, 'NOMBRES', formatoceldagris)
                    ws.write(0, 4, 'TIPO CONTRATO', formatoceldagris)
                    ws.write(0, 5, 'UNIDAD ORGÁNICA', formatoceldagris)
                    ws.write(0, 6, 'DENOMINACION PUESTO', formatoceldagris)
                    ws.write(0, 7, 'REGIMEN LABORAL', formatoceldagris)
                    cont =8

                    for campo in campos:
                        ws.write(0, cont, str(campo),formatoceldagris)
                        cont += 1

                    fila_dato_persona = 2
                    fila_detalle_valor = 1
                    band = True

                    for contra in contratopersona:
                        ws.write('A%s' % fila_dato_persona,
                                 str(contra.codigocontrato if contra.codigocontrato else 'No existe registro'),
                                 formatoceldaleft)
                        ws.write('B%s' % fila_dato_persona,
                                 str(contra.fechainicio if contra.fechainicio else 'No existe registro'),
                                 formatoceldaleft)
                        ws.write('C%s' % fila_dato_persona,
                                 str(contra.persona.cedula if contra.persona.cedula else 'No existe registro'),
                                 formatoceldaleft)
                        ws.write('D%s' % fila_dato_persona,
                                 str(contra.persona if contra.persona else 'No existe registro'), formatoceldaleft)
                        ws.write('E%s' % fila_dato_persona,
                                 str(contra.contrato.tipo if contra.contrato.tipo else 'No existe registro'),
                                 formatoceldaleft)
                        ws.write('F%s' % fila_dato_persona,
                                 str(contra.unidadorganica if contra.unidadorganica else 'No existe registro'),
                                 formatoceldaleft)
                        ws.write('G%s' % fila_dato_persona,
                                 str(contra.denominacionpuesto if contra.denominacionpuesto else 'No existe registro'),
                                 formatoceldaleft)
                        ws.write('H%s' % fila_dato_persona,
                                 str(contra.contrato.regimenlaboral if contra.contrato.regimenlaboral else 'No existe registro'),
                                 formatoceldaleft)

                        detallecontrato = ContratoPersonaDetalle.objects.filter(contratopersona=contra, status=True).order_by('campos')

                        columna_detalle_valor = 8
                        for campo in campos:
                            if detallecontrato.filter(campos=campo.campos).exists():
                                deta_valor = detallecontrato.get(campos=campo.campos)
                                ws.write(fila_detalle_valor, columna_detalle_valor,
                                         str(deta_valor.valor if deta_valor.valor else 'No existe registro'),
                                         formatoceldaleft)
                            columna_detalle_valor += 1

                        fila_dato_persona += 1
                        fila_detalle_valor += 1
                        band = False

                    workbook.close()
                    output.seek(0)
                    filename = 'plantilla_%s.xlsx' #% (contra.contrato.regimenlaboral)
                    response = HttpResponse(output,
                                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = 'attachment; filename=%s' % filename
                    return response
                except Exception as ex:
                    pass

            elif action == 'cargar_regimen':
                try:
                    lista = []
                    regimenes = OtroRegimenLaboral.objects.all()
                    for regimen in regimenes:
                        if not buscar_dicc(lista, 'id', regimen.id):
                            lista.append({'id': regimen.id, 'nombre':regimen.nombre})
                    return JsonResponse({'result': 'ok', 'lista': lista})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})


            elif action == 'importaciones':
                try:
                    data['title'] = u'Importaciones'
                    search = None
                    ids = None
                    if 's' in request.GET:
                        search = request.GET['s']
                        importacion = ArchivoContrato.objects.filter(Q(contrato__descripcion__icontains=search) |
                                                                       Q(contrato__regimenlaboral__nombre__icontains=search)).order_by('-id')
                    elif 'id' in request.GET:
                        ids = request.GET['id']
                        importacion = ArchivoContrato.objects.filter(id=ids).order_by('-id')
                    else:
                        importacion = ArchivoContrato.objects.all().order_by('-id')
                    paging = MiPaginador(importacion, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['importaciones'] = page.object_list
                    return render(request, "th_contrato/importaciones.html", data)
                except Exception as ex:
                    pass

            elif action == 'buscardepartamento':
                try:
                    q = request.GET['q'].upper().strip()
                    if q == 'UNDEFINED':
                        departamentos = Departamento.objects.filter(status=True,integrantes__isnull=False).distinct()
                    else:
                        departamentos = Departamento.objects.filter(status=True,integrantes__isnull=False, nombre__unaccent__icontains=q).distinct()[:15]
                    data = {"result": "ok",
                            "results": [{"id": d.id, "name": d.nombre} for d in departamentos]}
                    return JsonResponse(data)
                except Exception as ex:
                    pass

            return HttpResponseRedirect(request.path)
        else:

            # data['campos'] = CamposContratos.objects.filter(status=True).order_by('-id')
            #data['modeloscontratos'] = Contratos.objects.filter(status=True, anio=anioselect).order_by('-id')

            data['title'] = u'Contratos'
            data['anios'] = anios = rango_anios()
            if 'anio' in request.GET:
                request.session['aniotipocontrato'] = int(request.GET['anio'])
            if 'aniotipocontrato' not in request.session:
                request.session['aniotipocontrato'] = anios[0]
            data['anioselect'] = anioselect = request.session['aniotipocontrato']
            url_vars = ''
            filtro = Q(status=True)
            search = None
            ids = None

            if 's' in request.GET:

                search = request.GET['s'].strip()
                ss = search.split(' ')
                if len(ss) == 1:
                    filtro = filtro & (Q(persona__nombres__icontains=search) |
                                       Q(persona__apellido1__icontains=search) |
                                       Q(persona__apellido2__icontains=search) |
                                       Q(persona__cedula__icontains=search))
                else:
                    filtro = filtro & ((Q(persona__nombres__icontains=ss[0]) & Q(persona__nombres__icontains=ss[1])) |
                                       (Q(persona__apellido1__icontains=ss[0]) &
                                       Q(persona__apellido2__icontains=ss[1]) )|
                                       Q(persona__cedula__icontains=ss[0]))

            #     if request.GET['s'] != '':
            #         search = request.GET['s']
            #
            # if search:
            #     filtro = filtro & (Q(persona__nombres__icontains=search) |
            #                        Q(persona__apellido1__icontains=search) |
            #                        Q(persona__apellido2__icontains=search) |
            #                        Q(persona__cedula__icontains=search) )
                url_vars += '&s=' + search


            contratos = PersonaContratos.objects.filter(filtro).order_by('-id')
            # contratos = PersonaAplicarPartida.objects.filter(esganador=True,status=True)

            paging = MiPaginador(contratos, 20)
            p = 1
            try:
                paginasesion = 1
                if 'paginador' in request.session:
                    paginasesion = int(request.session['paginador'])
                if 'page' in request.GET:
                    p = int(request.GET['page'])
                    p = int(request.GET['page'])
                else:
                    p = paginasesion
                try:
                    page = paging.page(p)
                except:
                    p = 1
                page = paging.page(p)
            except:
                page = paging.page(p)
            request.session['paginador'] = p
            data['paging'] = paging
            data['rangospaging'] = paging.rangos_paginado(p)
            data['page'] = page
            data['search'] = search if search else ""
            data["url_vars"] = url_vars
            data['ids'] = ids if ids else ""
            data['contratos'] = page.object_list
            return render(request, 'th_contrato/view.html', data)
