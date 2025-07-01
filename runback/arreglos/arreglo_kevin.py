import os
import statistics
import sys
import xlwt
import openpyxl
import xlsxwriter
from xlwt import *

# SITE_ROOT = os.path.dirname(os.path.realpath(__file__))

YOUR_PATH = os.path.dirname(os.path.realpath(__file__))
# print(f"YOUR_PATH: {YOUR_PATH}")
SITE_ROOT = os.path.dirname(os.path.dirname(YOUR_PATH))
SITE_ROOT = os.path.join(SITE_ROOT, '')
# print(f"SITE_ROOT: {SITE_ROOT}")
your_djangoproject_home = os.path.split(SITE_ROOT)[0]
# print(f"your_djangoproject_home: {your_djangoproject_home}")
sys.path.append(your_djangoproject_home)

from webpush import send_user_notification
from django.core.wsgi import get_wsgi_application

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "settings")
application = get_wsgi_application()
from settings import SITE_STORAGE, MEDIA_ROOT, MEDIA_URL
from sga.models import *

from django.db import transaction, connections

from sga.funciones import null_to_decimal


def procesar_cierre_ingles_por_id_materia_asignada(id_materia_asignada):
    for materiaasignada in MateriaAsignada.objects.filter(pk__in=id_materia_asignada):
        materiaasignada.importa_nota = True
        materiaasignada.cerrado = True
        materiaasignada.fechacierre = datetime(2022, 12, 12, 0, 0, 0).date()
        materiaasignada.save()
        d = locals()
        exec(materiaasignada.materia.modeloevaluativo.logicamodelo, globals(), d)
        d['calculo_modelo_evaluativo'](materiaasignada)
        materiaasignada.cierre_materia_asignada()
        print(u"CIERRA -- %s" % (materiaasignada))


def proceso_cierre_ingles(periodo, id_nivel=1501):
    from django.http import HttpResponse
    response = HttpResponse(content_type="application/ms-excel")
    response['Content-Disposition'] = 'attachment; filename=reporte_notas_ingles.xls'
    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
    title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
    style1 = easyxf(num_format_str='D-MMM-YY')
    font_style = XFStyle()
    font_style.font.bold = True
    font_style2 = XFStyle()
    font_style2.font.bold = False
    wb = xlwt.Workbook()
    ws = wb.add_sheet('Sheetname')
    estilo = xlwt.easyxf(
        'font: height 350, name Arial, colour_index black, bold on, italic on; align: wrap on, vert centre, horiz center;')
    ws.write_merge(0, 0, 0, 9, 'UNIVERSIDAD ESTATAL DE MILAGRO', estilo)
    output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
    nombre = "Lista" + datetime.now().strftime('%Y%m%d_%H%M%S') + ".xls"
    filename = os.path.join(output_folder, nombre)
    columns = [(u"CEDULA", 6000),
               (u"APELLIDOS Y NOMBRES", 6000),
               (u"CARRERA", 6000),
               (u"URL", 6000),
               (u"MODULO", 6000),
               (u"NOTA BUCKINGHAM", 6000),
               (u"NOTA MATERIA", 6000),
               (u"ESTADO", 6000),
               (u"CURSO", 6000),
               (u"TIENE DEUDA", 6000),
               (u"VALOR", 6000)
               ]
    row_num = 3
    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num][0], font_style)
        ws.col(col_num).width = columns[col_num][1]
    row_num = 4
    cont = 0
    materiasasignadas = MateriaAsignada.objects.filter(status=True, materia__inglesepunemi=True, matricula__status=True,cerrado=False,
                                                       retiramateria=False, materia__nivel__periodo=periodo,
                                                       materia__nivel_id=id_nivel, materia__status=True,
                                                       importa_nota=False).order_by('matricula__inscripcion__carrera')
    print(u"%s" % (materiasasignadas.values('id').count()))
    for materiaasignada in materiasasignadas:
        with transaction.atomic():
            try:
                idcursomoodle = materiaasignada.materia.idcursomoodle
                url = 'https://upei.buckcenter.edu.ec/usernamecoursetograde.php?username=%s&curso=%s' % (
                materiaasignada.matricula.inscripcion.persona.identificacion(), idcursomoodle)
                cont += 1
                ws.write(row_num, 0, materiaasignada.matricula.inscripcion.persona.identificacion())
                ws.write(row_num, 1,
                         materiaasignada.matricula.inscripcion.persona.apellido1 + ' ' + materiaasignada.matricula.inscripcion.persona.apellido2 + ' ' + materiaasignada.matricula.inscripcion.persona.nombres)
                ws.write(row_num, 2, str(materiaasignada.matricula.inscripcion.carrera))
                ws.write(row_num, 3, str(url))
                req = Request(url)
                response = urlopen(req)
                result = json.loads(response.read().decode())
                idcurso = int(result['idcurso'])
                print(u"----- %s -----" % cont)
                print(u"PROCESANDO - %s" % materiaasignada)
                print(u"%s" % result)
                print(u"ID CURSO: %s" % idcurso)
                valores = 0
                rubros = materiaasignada.rubro.filter(status=True, observacion='INGLÉS ABRIL - AGOSTO 2023')
                for rubro in rubros:
                    valores = rubro.total_pagado()
                #         tiene_pagos=False
                # if tiene_pagos:
                if idcurso != idcursomoodle:
                    ws.write(row_num, 4, u"%s" % materiaasignada.materia)
                    ws.write(row_num, 5, u"NO COINCIDE CURSO")
                    ws.write(row_num, 6, u"NO COINCIDE CURSO")
                    ws.write(row_num, 7, u"NO COINCIDE CURSO")
                    ws.write(row_num, 8, u"NO COINCIDE CURSO")
                    ws.write(row_num, 9, u"NO COINCIDE CURSO")
                    ws.write(row_num, 10, u"NO COINCIDE CURSO")
                else:
                    try:
                        nota = null_to_decimal(result['nota'], 0)
                    except:
                        if result['nota'] == '-' or result['nota'] == None:
                            nota = 0

                    if nota != materiaasignada.notafinal and type(nota) in [int, float]:
                        campo = materiaasignada.campo('EX')
                        actualizar_nota_planificacion(materiaasignada.id, 'EX', nota)
                        auditorianotas = AuditoriaNotas(evaluaciongenerica=campo, manual=False, calificacion=nota)
                        auditorianotas.save()
                        materiaasignada.importa_nota = True
                        materiaasignada.cerrado = True
                        materiaasignada.fechacierre = datetime.now().date()
                        materiaasignada.save()
                        d = locals()
                        exec(materiaasignada.materia.modeloevaluativo.logicamodelo, globals(), d)
                        d['calculo_modelo_evaluativo'](materiaasignada)
                        materiaasignada.cierre_materia_asignada()
                        print(u"IMPORTA Y CIERRA -- %s" % (materiaasignada))
                    ws.write(row_num, 4, u"%s" % materiaasignada.materia)
                    ws.write(row_num, 5, u"%s" % result['nota'])
                    ws.write(row_num, 6, nota)
                    ws.write(row_num, 7, u"APROBADO" if nota >= 70 else "REPROBADO")
                    ws.write(row_num, 8,
                             u"COINCIDE CURSO" if idcurso == materiaasignada.materia.idcursomoodle else "NO COINCIDE CURSO")
                    ws.write(row_num, 9, u"TIENE RUBROS %s" % rubros.count() if rubros else 0)
                    ws.write(row_num, 10, u"VALORES %s" % valores if rubros else 0)

                    # if not MateriaAsignadaRetiro.objects.filter(status=True, materiaasignada=materiaasignada).exists():
                    #     retiro = MateriaAsignadaRetiro(materiaasignada=materiaasignada,
                    #                                    motivo='RETIRO POR TÉRMINO DEl PROCESO DE INGLÉS 1S 2023 BUCKINGHAM',
                    #                                    valida=False,
                    #                                    fecha=datetime.now().date())
                    #     retiro.save()
                    # if not materiaasignada.retiramateria:
                    #     materiaasignada.retiramateria = True
                    #     materiaasignada.save()
                    # rubros=materiaasignada.rubro.filter(status=True,observacion='INGLÉS REGULAR 2023 AGOSTO 2023')
                    # for rubro in rubros:
                    #     if not rubro.pagos():
                    #         rubro.delete()
                    #         print(u"ELIMINADO -- %s" % (materiaasignada))
                    # ws.write(row_num, 3, u"%s" % materiaasignada.materia)
                    # ws.write(row_num, 4, u"%s" % result['nota'])
                    # ws.write(row_num, 5, nota)
                    # ws.write(row_num, 6, u"RETIRADO")
                    # ws.write(row_num, 7,u"COINCIDE CURSO" if idcurso == materiaasignada.materia.idcursomoodle else "NO COINCIDE CURSO")
            except Exception as ex:
                transaction.set_rollback(True)
                print('error: %s' % (ex))
                ws.write(row_num, 4, u"%s" % ex)
                ws.write(row_num, 5, u"%s" % ex)
                ws.write(row_num, 6, u"%s" % ex)
                ws.write(row_num, 7, u"%s" % ex)
                ws.write(row_num, 8, u"%s" % ex)
                ws.write(row_num, 9, u"%s" % ex)
                ws.write(row_num, 10, u"%s" % ex)
                pass
            row_num += 1
    wb.save(filename)
    print("FIN: ", filename)



def query_modulos_ingles_por_ver_y_vistos(id_periodo):
    sql = f"""
    SELECT 
 coordinacion.nombre AS facultad,
 carrera.nombre AS carrera,
 (
SELECT nivelmalla.nombre
FROM sga_inscripcionnivel inscripcionnivel
INNER JOIN sga_nivelmalla nivelmalla ON inscripcionnivel.nivel_id = nivelmalla.id
WHERE inscripcionnivel.inscripcion_id = inscripcion.id) AS nivel,
 (persona.apellido1||' '||persona.apellido2||' '||persona.nombres) AS estudiante,
 persona.cedula,
 persona.email,
 persona.emailinst, 
 (
SELECT COUNT(asignatura.id)
FROM sga_malla malla
INNER JOIN sga_asignaturamalla asignaturamalla ON asignaturamalla.malla_id = malla.id
INNER JOIN sga_asignatura asignatura ON asignaturamalla.asignatura_id = asignatura.id
WHERE malla.id = 353 AND asignatura."status") AS num_modulos_ingles_malla,

 (
SELECT COUNT(recordacademico.id)
FROM sga_recordacademico recordacademico
INNER JOIN sga_asignatura asignatura0 ON recordacademico.asignatura_id = asignatura0.id
WHERE recordacademico.aprobada = TRUE AND recordacademico."status" AND asignatura0."status" AND recordacademico.inscripcion_id=inscripcion.id AND asignatura0.nombre ILIKE '%INGL%' AND asignatura0.id IN (
SELECT asignatura8.id
FROM sga_malla malla8
INNER JOIN sga_asignaturamalla asignaturamalla8 ON asignaturamalla8.malla_id = malla8.id
INNER JOIN sga_asignatura asignatura8 ON asignaturamalla8.asignatura_id = asignatura8.id
WHERE malla8.id = 353 AND asignatura8."status")) AS num_modulos_aprobados,

 (ARRAY_TO_STRING(array(
SELECT asignatura0.nombre
FROM sga_recordacademico recordacademico
INNER JOIN sga_asignatura asignatura0 ON recordacademico.asignatura_id = asignatura0.id
WHERE recordacademico.aprobada = TRUE AND recordacademico."status" AND asignatura0."status" AND recordacademico.inscripcion_id=inscripcion.id AND asignatura0.nombre ILIKE '%INGL%' AND asignatura0.id IN (
SELECT asignatura8.id
FROM sga_malla malla8
INNER JOIN sga_asignaturamalla asignaturamalla8 ON asignaturamalla8.malla_id = malla8.id
INNER JOIN sga_asignatura asignatura8 ON asignaturamalla8.asignatura_id = asignatura8.id
WHERE malla8.id = 353 AND asignatura8."status")
ORDER BY asignatura0.nombre),', ')) AS nombres_modulos_aprobados, 


 (
SELECT COUNT(asignatura8.id)
FROM sga_malla malla8
INNER JOIN sga_asignaturamalla asignaturamalla8 ON asignaturamalla8.malla_id = malla8.id
INNER JOIN sga_asignatura asignatura8 ON asignaturamalla8.asignatura_id = asignatura8.id
WHERE malla8.id = 353 AND asignatura8."status" AND asignatura8.id NOT IN (
SELECT asignatura1.id
FROM sga_recordacademico recordacademico1
INNER JOIN sga_asignatura asignatura1 ON recordacademico1.asignatura_id = asignatura1.id
WHERE recordacademico1.aprobada = TRUE AND recordacademico1."status" AND recordacademico1.inscripcion_id=inscripcion.id AND asignatura1.nombre ILIKE '%INGL%')) AS num_modulos_por_aprobar,

 (ARRAY_TO_STRING(array(
SELECT asignatura8.nombre
FROM sga_malla malla8
INNER JOIN sga_asignaturamalla asignaturamalla8 ON asignaturamalla8.malla_id = malla8.id
INNER JOIN sga_asignatura asignatura8 ON asignaturamalla8.asignatura_id = asignatura8.id
WHERE malla8.id = 353 AND asignatura8."status" AND asignatura8.id NOT IN (
SELECT asignatura1.id
FROM sga_recordacademico recordacademico1
INNER JOIN sga_asignatura asignatura1 ON recordacademico1.asignatura_id = asignatura1.id
WHERE recordacademico1.aprobada = TRUE AND recordacademico1."status" AND recordacademico1.inscripcion_id=inscripcion.id AND asignatura1.nombre ILIKE '%INGL%')
ORDER BY asignatura8.nombre),', ')) AS nombres_modulos_por_aprobar 
FROM sga_matricula matricula
INNER JOIN sga_inscripcion inscripcion ON matricula.inscripcion_id = inscripcion.id
INNER JOIN sga_inscripcionmalla inscripcionmalla ON inscripcionmalla.inscripcion_id = inscripcion.id
INNER JOIN sga_persona persona ON inscripcion.persona_id = persona.id
INNER JOIN sga_carrera carrera ON inscripcion.carrera_id= carrera.id
INNER JOIN sga_nivel nivel ON matricula.nivel_id = nivel.id
INNER JOIN sga_coordinacion_carrera coorcar ON coorcar.carrera_id = carrera.id
INNER JOIN sga_coordinacion coordinacion ON coorcar.coordinacion_id = coordinacion.id
WHERE nivel.periodo_id = {id_periodo} AND coordinacion.id IN (1,2,3,4,5)  and matricula."status" AND inscripcion."status"
 AND matricula.retiradomatricula = FALSE AND carrera."status"
ORDER BY coordinacion.nombre, carrera.nombre
    """
    return sql


def traer_datos_query(sql):
    try:
        cursor = connections['default'].cursor()
        cursor.execute(sql)
        rows_effected = cursor.rowcount
        listado = cursor.fetchall()
        campos = [desc[0] for desc in cursor.description]
        directory = os.path.join(MEDIA_ROOT, 'reportes', 'gestion')
        try:
            os.stat(directory)
        except:
            os.mkdir(directory)
        name_document = 'reporte'
        nombre_archivo = name_document + "_1.xlsx"
        directory = os.path.join(MEDIA_ROOT, 'reportes', 'gestion', nombre_archivo)

        _author_ = 'Unemi'
        workbook = xlsxwriter.Workbook(directory, {'constant_memory': True})
        ws = workbook.add_worksheet('resultados')
        fuentecabecera = workbook.add_format({
            'align': 'center',
            'bg_color': 'silver',
            'border': 1,
            'bold': 1
        })

        formatoceldacenter = workbook.add_format({
            'border': 1,
            'valign': 'vcenter',
            'align': 'center'})

        row_num, numcolum = 0, 0

        for col_num in campos:
            ws.write(row_num, numcolum, col_num, fuentecabecera)
            ws.set_column(row_num, numcolum, 40)
            numcolum += 1
        row_num += 1
        for lis in listado:
            colum_num = 0
            for l in lis:
                ws.write(row_num, colum_num, l, formatoceldacenter)
                ws.set_column(row_num, numcolum, 40)
                colum_num += 1
            row_num += 1

        workbook.close()
        response = HttpResponse(directory,
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=%s' % name_document
        #
        url_file = "{}reportes/gestion/{}".format(MEDIA_URL, nombre_archivo)
        print(url_file)

    except Exception as ex:
        print(ex)


#
# sql = query_modulos_ingles_por_ver_y_vistos(177)
# traer_datos_query(sql)


def borrar_materias_nivel_ingles():
    print("inicio.....")
    eMaterias = Materia.objects.filter(nivel_id=1508)
    print(f"total materias a borrar {eMaterias.count()}")
    for materia in eMaterias:
        print(f"borrando.. {materia}")
        materia.delete()
        print("borrado")
    print("fin")


#periodo=Periodo.objects.get(id=177)
#proceso_cierre_ingles(periodo)


def migrar_etapas_tutorias_por_mecanismo_titulacion_posgrado():
    print("inicio migracion tutorias posgrado")
    #obtengo todas las convocatorias que tienen configurado etapas generales
    convocatoria_id = ProgramaEtapaTutoriaPosgrado.objects.filter(status=True).values('convocatoria').order_by('convocatoria').distinct()
    eConfiguracionTitulacionPosgrados = ConfiguracionTitulacionPosgrado.objects.filter(status=True, id__in =convocatoria_id).distinct()
    #recorro convocatorias configuradas las etapas
    for eConfiguracionTitulacionPosgrado in eConfiguracionTitulacionPosgrados:
        #obtengo las etapas generales que se le configuraron a las convocatorias
        eProgramaEtapaTutoriaPosgradoAntiguo = ProgramaEtapaTutoriaPosgrado.objects.filter(status=True,convocatoria = eConfiguracionTitulacionPosgrado,mecanismotitulacionposgrado__isnull=True).order_by('orden')
        #recorro los mecanismo que se encuentran configurado en las convocatorias
        for mecanismo in eConfiguracionTitulacionPosgrado.carrera.get_mecanismo_configurados():
            #pregunto si no tiene configurado etapas le creo las etapas para ese mecanismo
            if not mecanismo.mecanismotitulacionposgrado.tiene_etapas_tutorias_configurado(eConfiguracionTitulacionPosgrado.pk):
                #recorro las etapas generales que se configuraron en la convocatoria para migrar a cada mecanismo
                for etapas_antiguas in eProgramaEtapaTutoriaPosgradoAntiguo:
                    eProgramaEtapaTutoriaPosgrado = ProgramaEtapaTutoriaPosgrado(
                        mecanismotitulacionposgrado =  mecanismo.mecanismotitulacionposgrado,
                        convocatoria = eConfiguracionTitulacionPosgrado,
                        etapatutoria = etapas_antiguas.etapatutoria,
                        orden = etapas_antiguas.orden,
                    )
                    eProgramaEtapaTutoriaPosgrado.save()

    print("fin migracion tutorias posgrado")



# migrar_etapas_tutorias_por_mecanismo_titulacion_posgrado()

def matricular_estudiantes_regulares(ePeriodoActual,ePeriodoAnterior):
    # ESTUDIANTES REGULARES : APROBARON TODAS LAS MATERIAS ANTERIORES
    # SI VOY A 4TO TODAS LAS DE 3RO DEBEN ESTAR APROBADAS Y NO TENER MATERIAS DE OTROS NIVELES MAYORES AL QUE VA
    # SI EL NIVEL AL QUE VOY TIENE MATERIAS DE PRACTICA EXCLUYO A ESA CARRERA
    ePeriodoMatricula = ePeriodoActual.periodomatricula_set.filter(status=True).first()
    ePersona = Persona.objects.get(pk=38453)
    usernotify = ePersona.usuario
    eNotificacion = Notificacion(cuerpo=f'{ePeriodoMatricula}',
                                 titulo=f'(En proceso) {ePeriodoMatricula}',
                                 destinatario=ePersona,
                                 url='',
                                 prioridad=1,
                                 app_label='SGA',
                                 fecha_hora_visible=datetime.now() + timedelta(days=5),
                                 tipo=2,
                                 en_proceso=True)

    EJE_FORMATIVO_PRACTICAS = 9
    EJE_FORMATIVO_VINCULACION = 11
    EXCLUDE_EJE_FORMATIVO = [EJE_FORMATIVO_PRACTICAS, EJE_FORMATIVO_VINCULACION]


    if not ePeriodoMatricula:
        raise NameError(u"Periodo académico de matricula no existe")

    print(f"PERIODO ACTUAL: {ePeriodoMatricula}")
    print(f"PERIODO ANTERIOR: {ePeriodoAnterior}")
    eCoordinaciones = Coordinacion.objects.filter(id__in=[1, 2, 3, 4, 5])
    eMatriculasAnterior = Matricula.objects.filter(status=True, retiradomatricula=False, bloqueomatricula=False, nivel__periodo=ePeriodoAnterior, inscripcion__carrera__coordinacion__in=eCoordinaciones)

    eMatriculasAnterior = eMatriculasAnterior.filter(cerrada=True)

    total_matriculas = eMatriculasAnterior.values("id").distinct().count()

    print(total_matriculas)
    contador_matricula = 0
    resultados = []
    errores =[]
    for eMatriculaAnterior in eMatriculasAnterior.filter(inscripcion__persona__cedula__in=['0940091499','0951358555']):
        with transaction.atomic():
            eInscripcion = eMatriculaAnterior.inscripcion
            if eInscripcion.itinerario > 0:
                print(f"Tiene itinerario: {eInscripcion.itinerario}")
            eInscripcionMalla = eInscripcion.malla_inscripcion()
            eMalla = eInscripcionMalla.malla
            eInscripcionNivel = eInscripcion.mi_nivel()
            itinerario = []
            itinerario.append(0)
            eAsignaturasMalla = eMalla.asignaturamalla_set.select_related().filter(status=True).exclude(Q(nivelmalla_id=NIVEL_MALLA_CERO) | Q(opcional=True) | Q(ejeformativo_id__in=EXCLUDE_EJE_FORMATIVO)  ).order_by('nivelmalla', 'ejeformativo').filter(itinerario__in= itinerario)

            if eInscripcion.itinerario>0:
                itinerario.append(eInscripcion.itinerario)
                eAsignaturasMalla = eAsignaturasMalla.filter(itinerario__in= itinerario )
                print(F"{eInscripcion.persona.cedula} - itinerario: {eInscripcion.itinerario}")
                print("_-----")

            eModulosMalla = ModuloMalla.objects.filter(malla=eInscripcionMalla.malla)
            eNivelMallaAprobadoActual = eInscripcionNivel.nivel
            eNivelMallaMatriculaAnterior = eMatriculaAnterior.nivelmalla
            eAsignaturasMallaquedebeaprobar = eAsignaturasMalla.filter(nivelmalla__orden__lte=eNivelMallaAprobadoActual.orden).exclude(Q(asignatura__id__in=eModulosMalla.values_list('asignatura__id', flat=True)) | Q( asignatura__in=AsignaturaMalla.objects.values_list("asignatura__id", flat=True).filter(malla = eMalla)))
            eAsignaturasRecord = eInscripcion.recordacademico_set.filter(asignaturamalla__in=eAsignaturasMalla,status=True)


            if eAsignaturasRecord.values("id").filter(aprobada=False).exists():
                errores.append("Tiene asignaturas en el record reprobadas")

            eAsignaturasAprobadashastanivel = eAsignaturasRecord.filter(asignaturamalla__in=eAsignaturasMallaquedebeaprobar, aprobada=True, status=True)

            if eAsignaturasAprobadashastanivel.values("id").count() < eAsignaturasMallaquedebeaprobar.values("id").count():
                errores.append("Tiene asignaturas en el record reprobadas")

            eAsignaturasAprobadasAdelantada = eAsignaturasRecord.filter(asignaturamalla__in=eAsignaturasMalla.filter(nivelmalla__orden__gte=eNivelMallaAprobadoActual.orden + 1).exclude(Q(asignatura__id__in=eModulosMalla.values_list('asignatura__id', flat=True)) | Q(asignatura__in=AsignaturaMalla.objects.values_list("asignatura__id", flat=True).filter(malla = eMalla))), aprobada=True, status=True)
            if eAsignaturasAprobadasAdelantada.values("id").exists():
                errores.append("Tiene asignaturas en el record que ha adelantado")

            # -----------------------------------------------------Verificar si hay asignaturas sin aprobar hasta el nivel actual
            eAsignaturasMallaHastaNivelActualTodas = eMalla.asignaturamalla_set.select_related().filter(status=True,itinerario=0, nivelmalla__orden__lte=eNivelMallaAprobadoActual.orden).exclude(Q(nivelmalla_id=NIVEL_MALLA_CERO) | Q(opcional=True) | Q(ejeformativo_id__in=EXCLUDE_EJE_FORMATIVO)  )
            asignaturas_sin_aprobar = eAsignaturasMallaHastaNivelActualTodas.exclude(id__in=eAsignaturasRecord.values_list("asignaturamalla_id",flat=True))
            if asignaturas_sin_aprobar.exists():
                errores.append(f"Tiene asignaturas sin aprobar hasta el nivel actual {eNivelMallaAprobadoActual.orden}")

            if eAsignaturasRecord.values("id").filter(aprobada=False).exists():
                errores.append("Tiene asignaturas en el record reprobadas")

                # -------------------



            eMateriaAsignadasAnterior = MateriaAsignada.objects.filter(matricula=eMatriculaAnterior, status=True).exclude( Q(materia__inglesepunemi=True) | Q( materia__asignatura__id__in=eModulosMalla.values_list('asignatura__id', flat=True)) | Q( materia__asignatura__in=AsignaturaMalla.objects.values_list("asignatura__id", flat=True).filter(malla =eMalla)))
            if eMateriaAsignadasAnterior.values("id").filter(Q(cerrado=False) | Q(estado_id__in=[2, 3, 4]) | Q(retiromanual=True)).exists():
                errores.append("Tiene asignaturas en la matricula anterior con estado cerrado=False | estado en [EN CURSO, REPROBADO] | retiromanual=True")

            ids_paralelo = eMateriaAsignadasAnterior.values_list('materia__paralelomateria__id', flat=True).distinct()
            if len(ids_paralelo) > 1:
                errores.append(f"Tiene materias con paralelos de más de {len(ids_paralelo)}")

            if Matricula.objects.values("id").filter(inscripcion=eInscripcion, nivel__periodo=ePeriodoActual).exists():
                errores.append(f"Ya existe una matricula en el periodo {ePeriodoActual.__str__()}")

            eAsignaturasMallaNivelSiguiente = eAsignaturasMalla.filter(nivelmalla__orden=eNivelMallaAprobadoActual.orden + 1)
            if eAsignaturasMallaNivelSiguiente.values("id").filter(practicas=True).exists():
                errores.append(f"Existe asignarturas de practicas en el nivel de malla {eNivelMallaAprobadoActual.orden + 1} a matricular")

            if eAsignaturasMallaNivelSiguiente.values("id").filter(itinerario__gt=0).exists():
                errores.append( f"Existe asignarturas con itinerario en el nivel de malla {eNivelMallaAprobadoActual.orden + 1} a matricular")
            eNivel = Nivel.objects.filter(periodo=ePeriodoActual,
                                          nivellibrecoordinacion__coordinacion__carrera=eInscripcion.carrera,
                                          nivellibrecoordinacion__coordinacion__sede=eInscripcion.sede,
                                          sesion=eMatriculaAnterior.nivel.sesion, cerrado=False,
                                          fin__gte=datetime.now().date()).first()
            if not eNivel:
                errores.append(f"No existe nivel en el periodo {ePeriodoActual.__str__()}")

            eParalelo = Paralelo.objects.filter(pk__in=ids_paralelo).first()
            eMateriasAbiertas = Materia.objects.filter(asignaturamalla__in=eAsignaturasMallaNivelSiguiente,
                                                       inicio__gte=datetime.now(), nivel__cerrado=False,
                                                       nivel__periodo=eNivel.periodo, paralelomateria=eParalelo,
                                                       status=True).order_by('id')
            eMateriasAbiertas = eMateriasAbiertas.annotate(numasignadados=Count('materiaasignada__id', distinct=True)).filter(cupo__gt=F('numasignadados'))
            if not eMateriasAbiertas.values("id").exists():
                errores.append(f"No existe materias aperturadas del paralelo ({eParalelo.__str__()}) para el periodo {ePeriodoActual.__str__()}")
            if eAsignaturasMallaNivelSiguiente.values("id").count() != eMateriasAbiertas.values("id").count():
                errores.append(f"No existe suficiente materias aperturadas para el periodo {ePeriodoActual.__str__()}")

            if errores:
                print("---------------------------------")
                print(f"{eMatriculaAnterior} - {eNivelMallaAprobadoActual}")
                for error in errores:
                    print(error)
                errores.clear()
                print("---------------------------------")
            else:
                print(f"({total_matriculas}/{contador_matricula}) - {eMatriculaAnterior.__str__()}")

                #AQUI inicio MATRICULA .---------------------------------

                # AQUI fin MATRICULA .---------------------------------
                resultados.append({'eInscripcion': eInscripcion,
                                   'eNivelMallaAprobadoActual': eNivelMallaAprobadoActual,
                                   'eMatriculaAnterior': eMatriculaAnterior,
                                   'eParaleloAnterior': eParalelo,
                                   'eMateriaAsignadasAnterior': eMateriaAsignadasAnterior,
                                   'eAsignaturasMallaNivelSiguiente': eAsignaturasMallaNivelSiguiente,
                                   # 'eMatriculaActual': eMatriculaActual,
                                   # 'eMateriaAsignadasActual': eMateriaAsignadasActual,
                                   # 'eParaleloActual': eParaleloActual,
                                   })
            itinerario.clear()

    if len(resultados) > 0:
            nombre_archivo = "reporte_automaricula_2s_2024"
            # fuentecabecera = easyxf('font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
            # fuentenormal = easyxf('font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
            style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
            style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
            style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
            title = easyxf(
                'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
            style1 = easyxf(num_format_str='D-MMM-YY')
            font_style = XFStyle()
            font_style.font.bold = True
            font_style2 = XFStyle()
            font_style2.font.bold = False
            wb = xlwt.Workbook()
            ws = wb.add_sheet('Sheetname')
            estilo = xlwt.easyxf(
                'font: height 350, name Arial, colour_index black, bold on, italic on; align: wrap on, vert centre, horiz center;')
            output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
            nombre = nombre_archivo + "_" + datetime.now().strftime('%Y%m%d_%H%M%S') + ".xls"
            filename = os.path.join(output_folder, nombre)

            columnas = [
                (u"#", 7000),
                (u"documento", 7000),
                (u"alumno", 7000),
                (u"facultad", 7000),
                (u"carrera", 7000),
                (u"nivelmallaaprobado", 7000),
                (u"matriculaanterior", 7000),
                (u"nivelmallamatriculaanterior", 7000),
                (u"paraleloanterior", 7000),
                (u"materiasanterior", 7000),
                (u"asignaturasdebetomar", 7000),
                (u"matriculaactual", 7000),
                (u"materiasactual", 7000),
                (u"paraleloactual", 7000),
            ]
            row_num = 0
            for col_num in range(len(columnas)):
                ws.write(row_num, col_num, columnas[col_num][0], font_style)
                ws.col(col_num).width = columnas[col_num][1]
            total = len(resultados)
            cont = 0
            row_num = 1
            i = 0
            for r in resultados:
                cont += 1
                i += 1
                eInscripcion = r.get('eInscripcion')
                ePersona = eInscripcion.persona
                eNivelMallaAprobadoActual = r.get('eNivelMallaAprobadoActual')
                eMatriculaAnterior = r.get('eMatriculaAnterior')
                eParaleloAnterior = r.get('eParaleloAnterior')
                eMateriaAsignadasAnterior = r.get('eMateriaAsignadasAnterior')
                eAsignaturasMallaNivelSiguiente = r.get('eAsignaturasMallaNivelSiguiente')
                eMatriculaActual = r.get('eMatriculaActual')
                eMateriaAsignadasActual = r.get('eMateriaAsignadasActual')
                eParaleloActual = r.get('eParaleloActual')
                eFacultad = eInscripcion.coordinacion
                eCarrera = eInscripcion.carrera
                ws.write(row_num, 0, i, font_style2)  #
                ws.write(row_num, 1, "%s" % ePersona.documento(), font_style2)  # documento
                ws.write(row_num, 2, "%s (%s)" % (ePersona.nombre_completo_inverso(), eInscripcion.id),font_style2)  # alumno
                ws.write(row_num, 3, "%s" % eFacultad.__str__(), font_style2)  # facultad
                ws.write(row_num, 4, "%s" % eCarrera.__str__(), font_style2)  # carrera
                ws.write(row_num, 5, "%s" % eNivelMallaAprobadoActual.__str__(), font_style2)  # nivelmallaaprobado
                ws.write(row_num, 6, "ID_MATRICULA: %s - PERIODO: %s" % (eMatriculaAnterior.id, eMatriculaAnterior.nivel.periodo.__str__()), font_style2)  # matriculaanterior
                ws.write(row_num, 7, "%s" % eMatriculaAnterior.nivelmalla.__str__(),font_style2)  # nivelmallamatriculaanterior
                ws.write(row_num, 8, "%s" % eParaleloAnterior.__str__(), font_style2)  # paraleloanterior
                ws.write(row_num, 9,"%s" % ', '.join([x.materia.asignatura.nombre for x in eMateriaAsignadasAnterior]), font_style2)  # materiasanterior
                ws.write(row_num, 10, "%s" % ', '.join([x.asignatura.nombre for x in eAsignaturasMallaNivelSiguiente]), font_style2)  # asignaturasdebetomar
                # ws.write(row_num, 11, "ID_MATRICULA: %s - PERIODO: %s" % ( eMatriculaActual.id, eMatriculaActual.nivel.periodo.__str__()), font_style2)  # matriculaactual
                # ws.write(row_num, 12,"%s" % ', '.join([x.materia.asignatura.nombre for x in eMateriaAsignadasActual]),font_style2)  # materiasactual
                # ws.write(row_num, 13, "%s" % eParaleloActual.__str__(), font_style2)  # paraleloactual
                row_num += 1
                print(f"({total}/{cont}) Procesando . . .")
            wb.save(filename)
            print("ARCHIVO: ", filename)
            eNotificacion.url = "{}/{}".format(MEDIA_URL, nombre)
            eNotificacion.titulo = f'(Finalizado) Generación de automatricula 2S 2024 PREGRADO'
            eNotificacion.save()
            print("Proceso finalizado . . .")



#descomentar_enelreporte y en result: cuando se matriculen y    quitar  eMatriculasAnterior[600:800]:
ePeriodoActual = Periodo.objects.get(pk=336)
ePeriodoAnterior = Periodo.objects.get(pk=317)
matricular_estudiantes_regulares(ePeriodoActual,ePeriodoAnterior)
