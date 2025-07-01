# coding=utf-8
#!/usr/bin/env python

import os
import sys


# SITE_ROOT = os.path.dirname(os.path.realpath(__file__))
import xlrd
from openpyxl import load_workbook

YOUR_PATH = os.path.dirname(os.path.realpath(__file__))
# print(f"YOUR_PATH: {YOUR_PATH}")
SITE_ROOT = os.path.dirname(os.path.dirname(YOUR_PATH))
SITE_ROOT = os.path.join(SITE_ROOT, '')
# print(f"SITE_ROOT: {SITE_ROOT}")
your_djangoproject_home = os.path.split(SITE_ROOT)[0]
# print(f"your_djangoproject_home: {your_djangoproject_home}")
sys.path.append(your_djangoproject_home)

from django.core.wsgi import get_wsgi_application
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "settings")
application = get_wsgi_application()
from datetime import datetime, timedelta
from django.db import transaction
from sga.models import Materia, Clase, Leccion, LeccionGrupo, AsistenciaLeccion
from sga.funciones import variable_valor, validar_ldap_reseteo, convertir_hora
from sga.models import *
from sagest.models import *
from moodle.models import UserAuth
from datetime import datetime, timedelta
from django.db import transaction
from sga.models import Materia, Clase, Leccion, LeccionGrupo, AsistenciaLeccion, CamposTitulosPostulacion
from sga.funciones import variable_valor
# from sga.models import *
# from sagest.models import *
import xlrd
from postulate.models import Partida, PersonaAplicarPartida, PersonaFormacionAcademicoPartida, PartidaTribunal
from sga.models import *
from sagest.models import *
from django.db.models import Sum, F, FloatField
from django.db.models.functions import Coalesce
from settings import MEDIA_ROOT, BASE_DIR, CALCULO_POR_CREDITO
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, colors
from django.http import HttpResponse
from sga.models import Modulo
from gdocumental.models import *
from bd.models import *

try:
    archivo_ = 'admision1s2022_1'
    url_archivo = "{}/media/{}.xlsx".format(SITE_STORAGE, archivo_)
    workbook = load_workbook(filename=url_archivo, read_only=False)
    sheet = workbook[workbook.sheetnames[3]]
    linea = 1
    col_carrera = 1
    col_cedula = 3
    totalcambiados = 0
    totalexcluidos = 0
    periodo_ = Periodo.objects.get(pk=158)
    listexcluidossininscripcion = []
    listexcluidossinmatricula = []
    listadonoalcanzoex = []
    listbloqueado = []
    totalleer = sheet.max_row-1
    leido = 0
    for rowx in range(2, sheet.max_row + 1):
        if True:
            idcarrera_ = sheet.cell(row=rowx,column=col_carrera).value
            cedula_ = sheet.cell(row=rowx,column=col_cedula).value
            persona_ = Persona.objects.filter(cedula__icontains=cedula_, status=True, inscripcion__isnull=False).order_by('-id').first()
            carrera_ = Carrera.objects.get(id=idcarrera_)
            qsinscripcion = Inscripcion.objects.filter(status=True, persona=persona_, carrera=carrera_).order_by('-id').first()
            if qsinscripcion:
                qsmatricula = Matricula.objects.filter(inscripcion=qsinscripcion, nivel__periodo=periodo_).order_by('-id').first()
                if qsmatricula:
                    materias_ = MateriaAsignada.objects.filter(status=True, matricula=qsmatricula)
                    print(f'Procesando: {qsmatricula}')
                    if qsmatricula.bloqueomatricula:
                        listbloqueado.append(cedula_)
                    # for m in materias_:
                    #     notaanterior, notaactual = m.notafinal, m.notafinal
                    #     p = m.evaluaciongenerica_set.filter(status=True, detallemodeloevaluativo_id=115).first().valor
                    #     ex = m.evaluaciongenerica_set.filter(status=True, detallemodeloevaluativo_id=114).first().valor
                    #     n4 = m.evaluaciongenerica_set.filter(status=True, detallemodeloevaluativo_id=113).first().valor
                    #     n3 = m.evaluaciongenerica_set.filter(status=True, detallemodeloevaluativo_id=112).first().valor
                    #     n2 = m.evaluaciongenerica_set.filter(status=True, detallemodeloevaluativo_id=111).first().valor
                    #     n1 = m.evaluaciongenerica_set.filter(status=True, detallemodeloevaluativo_id=110).first().valor
                        # if m.notafinal < 70:
                        #     puntosfaltante = 70 - m.notafinal
                        #     if ex < 60:
                        #         puntosfaltantex = 60 - ex
                        #         if puntosfaltante <= puntosfaltantex:
                        #             notanuevaexamen = (ex + puntosfaltante)
                        #             notaactual = (notaanterior + puntosfaltante)
                        #             exfinal = m.evaluaciongenerica_set.filter(status=True, detallemodeloevaluativo_id=114).first()
                        #             exfinal.valor = notanuevaexamen
                        #             exfinal.save()
                        #             m.actualiza_notafinal()
                        #             print(f'----Materia {m.materia.__str__()}: {notaanterior} - {notaactual}')
                        #             print(f'----Ex: {ex} - {notanuevaexamen} (+{puntosfaltante})')
                        #         else:
                        #             notanuevaexamen = (ex + puntosfaltante)
                        #             notaactual = (notaanterior + puntosfaltante)
                        #             exfinal = m.evaluaciongenerica_set.filter(status=True, detallemodeloevaluativo_id=114).first()
                        #             exfinal.valor = notanuevaexamen
                        #             exfinal.save()
                        #             m.actualiza_notafinal()
                        #             print(f'----Materia {m.materia.__str__()}: {notaanterior} - {notaactual}')
                        #             print(f'----Ex: {ex} - {notanuevaexamen}')
                        #             listadonoalcanzoex.append(qsmatricula.id)
                        #             print('NO ALCANZA PUNTOS CON EXAMEN')
                    totalcambiados += 1
                else:
                    listexcluidossinmatricula.append(cedula_)
            else:
                listexcluidossininscripcion.append(cedula_)
            leido += 1
            print(f'-Leyendo {leido}/{totalleer}-')
    print('----------------------Sin Inscripcion-------------------')
    print(f'{listexcluidossininscripcion}')
    print('----------------------Sin Matricula---------------------')
    print(f'{listexcluidossinmatricula}')
    print('----------------------NO ALCANZO EXAMEN---------------------')
    print(f'{listadonoalcanzoex}')
    print('----------------------BLOQUEADO---------------------')
    print(f'{listbloqueado}')
    print('--------------------------------------------------------')
    print(f'Matriculados: {totalcambiados}')
    print(f'No alcanzaron con examen: {len(listadonoalcanzoex)}')
except Exception as ex:
    print(ex)
