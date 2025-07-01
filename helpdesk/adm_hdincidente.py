# -*- coding: UTF-8 -*-
import json
from datetime import datetime, timedelta
from django.contrib.auth.decorators import login_required
from django.forms import model_to_dict
import xlwt
from django.db import transaction
from django.db.models import Q, Sum, Count
from django.db.models.aggregates import Count
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.shortcuts import render
from django.template.context import Context
from django.template.loader import get_template
from openpyxl import Workbook
from rest_framework import reverse

from decorators import secure_module
from xlwt import *
import random

from posgrado.forms import RequisitosMaestriaForm
from helpdesk.commonviews import secuencia_ordentrabajo
from helpdesk.forms import HdGrupoForm, HdDetalle_GrupoForm, HdIncidenciaFrom, HdDetalleIncidenteFrom, HdEstadoForm, \
    HdEstadoEditForm, \
    HdEstadoImagenForm, HdImpactoForm, HdUrgenciaForm, HdPrioridadForm, HdMedioReporteForm, \
    HdUrgencia_Impacto_PrioridadForm, HdCatrgoriaForm, HdSubCatrgoriaForm, HdDetalleSubCategoriaForm, HdProcesoFrom, \
    HdEstadoProcesoFrom, HdDirectorForm, HdPrioridadImagenForm, HdTipoIncidenteForm, HdPiezaPartesForm, \
    HdSolicitudPiezaPartesForm, HdCausasForm, HdCabEncuestasForm, HdDetEncuestasForm, HdPreciosForm, HdFechacierreForm, \
    SeleccionarActivoForm, CerrarOrdenForm, DetalleOrdenForm, HdMaterialForm, HdMaterial_IncidenteForm, \
    HdIncidenciaLicenciaFrom, HdDetalleIncidenciaLicenciaFrom, HdDetalleIncidenciaProductoFrom, IncidenteForm
from helpdesk.models import HdGrupo, HdDetalle_Grupo, HdIncidente, HdSubCategoria, HdDetalle_SubCategoria, Departamento, \
    ActivoFijo, HdDetalle_Incidente, HdMaterial_Incidente, \
    HdEstado, HdImpacto, HdUrgencia, HdPrioridad, HdMedioReporte, HdUrgencia_Impacto_Prioridad, HdCategoria, HdProceso, \
    HdEstado_Proceso, HdDirector, HdTipoIncidente, HdPiezaPartes, \
    HdSolicitudesPiezaPartes, HdCausas, HdRequerimientosPiezaPartes, HdCabEncuestas, HdDetEncuestas, \
    HdRespuestaEncuestas, HdPrecioSolicitudesPiezaPartes, HdFechacierresolicitudes, \
    OrdenTrabajo, DetalleOrdenTrabajo, HdDetalle_Incidente_Ayudantes, HdMateriales, ESTADO_ORDEN_TRABAJO, \
    BodegaProducto, BodegaProductoDetalle, HdIncidenteProductoDetalle, actualizar_kardex_e_s, BodegaKardex
from sagest.models import Producto, TipoProducto, InformeActivoBaja, HdBloque, HdUbicacion, HdBloqueUbicacion, \
    DocumentoFirmaInformeBaja
from sga.commonviews import adduserdata, obtener_reporte
from sga.funciones import MiPaginador, log, generar_nombre, convertir_fecha, convertir_fecha_hora, notificacion2, notificacion, validar_archivo
from sga.models import Administrativo, Persona, MESES_CHOICES, SagPreguntaEncuesta
from sga.funcionesxhtml2pdf import add_tabla_reportlab, conviert_html_to_pdf, generar_pdf_reportlab, add_titulo_reportlab, add_graficos_barras_reportlab, add_graficos_circular_reporlab
from sga.templatetags.sga_extras import encrypt

# @login_required(redirect_field_name='ret', login_url='/loginsagest')
# @secure_module
@transaction.atomic()
def view(request):
    data = {}
    adduserdata(request, data)
    usuario = request.user
    persona = request.session['persona']
    hoy = datetime.now()
    if request.method == 'POST':
        action = request.POST['action']
        if action == 'add':
            try:
                form = HdIncidenciaFrom(request.POST, request.FILES)
                if not int(request.POST['persona']) > 0:
                    return JsonResponse({"result": "bad", "mensaje": u"No ha seleccionado quien solicinta el incidente."})
                newfile = None
                if 'archivo' in request.FILES:
                    newfile = request.FILES['archivo']
                    if newfile:
                        if newfile.size > 4194304:
                            return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 4 Mb."})
                        else:
                            newfilesd = newfile._name
                            ext = newfilesd[newfilesd.rfind("."):]
                            if ext == '.pdf' or ext == '.jpg' or ext == '.png' or ext == '.jpeg' or ext == '.PDF' or ext == '.JPG' or ext == '.PNG' or ext == '.JPEG':
                                newfile._name = generar_nombre("estadohelpdesk_", newfile._name)
                            else:
                                return JsonResponse({"result": "bad", "mensaje": u"Error, Solo archivo con extención. pdf, jpg, png, jpeg."})
                if form.is_valid():
                    if HdDirector.objects.filter(vigente=True).exists():
                        activo = None
                        responsableactivofijo = None
                        idcausa = None
                        getactivo = None
                        revisionequipoexterno = form.cleaned_data['revisionequipoexterno'] if 'revisionequipoexterno' in request.POST else False
                        revisionequiposincodigo = form.cleaned_data['revisionequiposincodigo'] if 'revisionequiposincodigo' in request.POST else False
                        if int(request.POST['tipoincidente']) == 2:
                            if not (revisionequipoexterno or revisionequiposincodigo):
                                if request.POST['activo'] == '':
                                    return JsonResponse({"result": "bad", "mensaje": u"No ha seleccionado el activo."})
                                if not int(request.POST['activo']) > 0:
                                    return JsonResponse({"result": "bad", "mensaje": u"No ha seleccionado el activo."})
                        if not (revisionequipoexterno or revisionequiposincodigo):
                            if len(request.POST['activo']) > 0:
                                if int(request.POST['activo']) > 0:
                                    getactivo = ActivoFijo.objects.get(pk=int(request.POST['activo']))
                                    if getactivo.responsable:
                                        responsableactivofijo = getactivo.responsable.id
                                    else:
                                        responsableactivofijo = None
                        if len(request.POST['causa']) > 0:
                            if int(request.POST['causa']) > 0:
                                idcausa = int(request.POST['causa'])

                        incidente = HdIncidente(asunto=form.cleaned_data['asunto'],
                                                persona_id=int(request.POST['persona']),
                                                ubicacion_id=int(request.POST['ubicacion']),
                                                descripcion=form.cleaned_data['descripcion'],
                                                subcategoria_id=request.POST['subcategoria'],
                                                detallesubcategoria_id=request.POST['detallesubcategoria'] if request.POST['detallesubcategoria'] else None,
                                                activo_id=getactivo.id if getactivo else getactivo,
                                                responsableactivofijo_id=responsableactivofijo,
                                                fechareporte=form.cleaned_data['fechareporte'],
                                                # fechareporte=datetime.now().date(),
                                                horareporte=form.cleaned_data['horareporte'],
                                                medioreporte_id=int(request.POST['medioreporte']),
                                                estado_id=int(request.POST['estado']),
                                                tipoincidente_id=int(request.POST['tipoincidente']),
                                                causa_id=idcausa,
                                                director_id=HdDirector.objects.get(vigente=True, status=True).persona.id,
                                                revisionequipoexterno=form.cleaned_data['revisionequipoexterno'] if 'revisionequipoexterno' in request.POST else False,
                                                revisionequiposincodigo=form.cleaned_data['revisionequiposincodigo'] if 'revisionequiposincodigo' in request.POST else False,
                                                serie=form.cleaned_data['serie'] if 'serie' in request.POST else '',
                                                archivo=newfile
                                                )
                        incidente.save(request)

                        # incidente.email_notificacion_tic(request.session['nombresistema'])
                        log(u'Adiciono nuevo Incidente: %s' % incidente, request, "add")
                        if request.POST['grupo']:
                            grupo = HdGrupo.objects.get(pk=int(request.POST['grupo']))
                            if HdDetalle_Grupo.objects.filter(responsable=True,grupo=grupo.pk).exists():
                                detalle = HdDetalle_Incidente(incidente_id=incidente.id,
                                                              grupo_id=int(request.POST['grupo']) if request.POST['grupo'] else None,
                                                              agente_id=request.POST['agente'] if request.POST['agente'] else None,
                                                              estadoasignacion=1,
                                                              resolucion=form.cleaned_data['resolucion'] if 'resolucion' in request.POST else None,
                                                              # fecharesolucion=form.cleaned_data['fecharesolucion'],
                                                              fecharesolucion=datetime.now().date(),
                                                              horaresolucion=datetime.now().time(),
                                                              estadoproceso_id=int(request.POST['estadobaja']) if 'estadobaja' in request.POST else None,
                                                              responsable_id=HdDetalle_Grupo.objects.get(responsable=True,grupo=grupo.pk).persona.id,
                                                              estado_id=int(request.POST['estado'])
                                                              )
                                detalle.save(request)
                                listaayudantes = HdDetalle_Grupo.objects.filter(pk__in=form.cleaned_data['ayudantes'])
                                if listaayudantes:
                                    for ayudante in listaayudantes:
                                        detalleayudantes = HdDetalle_Incidente_Ayudantes(detallleincidente=detalle,
                                                                                         agente=ayudante
                                                                                         )
                                        detalleayudantes.save(request)
                                if request.POST['agente']:
                                    detalle.email_notificacion_agente(request.session['nombresistema'])
                                log(u'Adiciono nuevo Detalle: %s' % detalle, request, "add")
                            else:
                                return JsonResponse({"result": "bad", "mensaje": u"No ha definido un Responsable, registreló en configuración, Agentes."})
                        if incidente.revisionequiposincodigo and incidente.tipoincidente.id == 2:
                            incidente.email_notificacion_equipo_sin_codigo(request.session['nombresistema'])
                        return JsonResponse({"result": "ok"})
                    else:
                        return JsonResponse({"result": "bad", "mensaje": u"No ha definido un Director,registrelo en configuración."})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'addincidente':
            try:
                f = IncidenteForm(request.POST, request.FILES)

                if 'archivo' in request.FILES:
                    archivo = request.FILES['archivo']
                    descripcionarchivo = 'Evidencia'
                    resp = validar_archivo(descripcionarchivo, archivo, ['pdf', 'jpg', 'jpeg', 'png'], '4MB')
                    if resp['estado'] != "OK":
                        return JsonResponse({"result": "bad", "titulo": "Atención!!!", "mensaje": resp["mensaje"], "showSwal": "True", "swalType": "warning"})

                if f.is_valid():
                    perfilprin = request.session['perfilprincipal']
                    archivo = None
                    if 'archivo' in request.FILES:
                        archivo = request.FILES['archivo']
                        archivo._name = generar_nombre("estadohelpdesk_", archivo._name)

                    # Consulto el activo
                    activofijo = ActivoFijo.objects.get(pk=f.cleaned_data['activo']) if f.cleaned_data['activo'] else None

                    # Guardo el incidente
                    incidente = HdIncidente(
                        asunto=f.cleaned_data['incidente'].strip(),
                        ubicacion=f.cleaned_data['ubicacion'],
                        persona_id=f.cleaned_data['persona'],
                        fechareporte=f.cleaned_data['fecha'],
                        horareporte=f.cleaned_data['hora'],
                        medioreporte_id=4,
                        estado_id=1,
                        tercerapersona=None,
                        tipoincidente=f.cleaned_data['tipogrupo'],
                        activo=activofijo,
                        responsableactivofijo=activofijo.responsable if activofijo else None,
                        archivo=archivo,
                        tipousuario=2 if perfilprin.es_profesor() else 1,
                        activosincodigo=None,
                        concodigo=True
                    )
                    incidente.save(request)

                    log(u'%s adicionó indicente: %s' % (persona, incidente), request, "add")
                    return JsonResponse({"result": "ok", "titulo": "Proceso exitoso!!!", "mensaje": u"Registro guardado con éxito", "showSwal": True})
                else:
                    for k, v in f.errors.items():
                        raise NameError( k + ', ' + v[0])
            except Exception as ex:
                msg = ex.__str__()
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al guardar los datos. [%s]" % msg, "showSwal": "True", "swalType": "error"})

        elif action == 'editincidente':
            try:
                # Consulto el incidente
                incidente = HdIncidente.objects.get(pk=int(encrypt(request.POST['id'])))

                f = IncidenteForm(request.POST, request.FILES)

                if 'archivo' in request.FILES:
                    archivo = request.FILES['archivo']
                    descripcionarchivo = 'Evidencia'
                    resp = validar_archivo(descripcionarchivo, archivo, ['pdf', 'jpg', 'jpeg', 'png'], '4MB')
                    if resp['estado'] != "OK":
                        return JsonResponse({"result": "bad", "titulo": "Atención!!!", "mensaje": resp["mensaje"], "showSwal": "True", "swalType": "warning"})

                if f.is_valid():
                    # Consulto el activo
                    activofijo = ActivoFijo.objects.get(pk=f.cleaned_data['activo']) if f.cleaned_data['activo'] else None

                    # Actualizo el incidente
                    incidente.asunto = f.cleaned_data['incidente'].strip()
                    incidente.ubicacion = f.cleaned_data['ubicacion']
                    incidente.persona_id = f.cleaned_data['persona']
                    incidente.fechareporte = f.cleaned_data['fecha']
                    incidente.horareporte = f.cleaned_data['hora']
                    incidente.tipoincidente = f.cleaned_data['tipogrupo']
                    incidente.activo = activofijo
                    incidente.responsableactivofijo = activofijo.responsable if activofijo else None

                    if 'archivo' in request.FILES:
                        archivo = request.FILES['archivo']
                        archivo._name = generar_nombre("estadohelpdesk_", archivo._name)
                        incidente.archivo = archivo

                    incidente.save(request)

                    log(u'%s editó incidente: %s' % (persona, incidente), request, "edit")
                    return JsonResponse({"result": "ok", "titulo": "Proceso exitoso!!!", "mensaje": u"Registro actualizado con éxito", "showSwal": True})
                else:
                    for k, v in f.errors.items():
                        raise NameError( k + ', ' + v[0])
            except Exception as ex:
                msg = ex.__str__()
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al guardar los datos. [%s]" % msg, "showSwal": "True", "swalType": "error"})

        elif action == 'anularresolucion':
            try:
                if 'id' not in request.POST:
                    return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al procesar la solicitud", "showSwal": "True", "swalType": "error"})

                # Consulto el incidente
                incidente = HdIncidente.objects.get(pk=int(encrypt(request.POST['id'])))

                # Verifico si se instaló licencia de software
                # 19 -> ANTIVIRUS SOPHOS
                sw_instalado = incidente.software_instalado([19])

                if sw_instalado is False:
                    # Actualizo el incidente
                    incidente.estado_id = 1
                    incidente.subcategoria = None
                    incidente.detallesubcategoria = None
                    incidente.causa = None
                    incidente.revisionequipoexterno = False
                    incidente.revisionequiposincodigo = False
                    incidente.serie = ''
                    incidente.observacion = request.POST['observacion'].strip()
                    incidente.save(request)
                else:
                    # Elimino el incidente
                    incidente.observacion = request.POST['observacion'].strip()
                    incidente.status = False
                    incidente.save(request)

                # Consulto y elimino el detalle (Resoluciones)
                for detalle in incidente.h_d_incidentes.filter(status=True).order_by('id'):
                    detalle.status = False
                    detalle.save(request)

                # Consulto detalle de productos utilizados, los elimino y creo el movimiento de kardex como ingreso
                for detalle in incidente.productos_utilizados():
                    # Eliminar producto
                    detalle.status = False
                    detalle.save(request)

                    # Consulto el registro más reciente de Kardex
                    kardexreciente = BodegaKardex.objects.filter(status=True, producto=detalle.producto).latest('fecha')
                    saldo_inicial = kardexreciente.saldoFinal
                    saldo_final = saldo_inicial + detalle.cantidad

                    nuevokardex = BodegaKardex(
                        bodega_id=1,
                        producto=detalle.producto,
                        tipotransaccion_id=1,
                        unidadmedida=detalle.unidadmedida,
                        cantidad=detalle.cantidad,
                        saldoInicial=saldo_inicial,
                        saldoFinal=saldo_final,
                        fecha=datetime.now(),
                        detalleincidente=detalle,
                        observacion='INGRESO POR MOTIVO DE ANULACIÓN DE INCIDENTE RESUELTO. ID INCIDENTE: {}'.format(incidente.id)
                    )
                    nuevokardex.save(request)

                log(u'%s anuló resolución del incidente: %s' % (persona, incidente), request, "edit")
                return JsonResponse({"result": "ok", "titulo": "Proceso exitoso!!!", "mensaje": u"Registro guardado con éxito", "showSwal": True})
            except Exception as ex:
                msg = ex.__str__()
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al guardar los datos. [%s]" % msg, "showSwal": "True", "swalType": "error"})

        elif action == 'addsalidalicenciasoft':
            try:
                from wpush.models import SubscriptionInfomation
                from webpush.utils import _send_notification
                form = HdIncidenciaLicenciaFrom(request.POST)
                # if not int(request.POST['persona']) > 0:
                #     return JsonResponse(
                #         {"result": "bad", "mensaje": u"No ha seleccionado quien solicinta el incidente."})

                if not request.POST.getlist('producto')[0]:

                    return JsonResponse({"result": "bad", "mensaje": u"Seleccione una Licencia"})

                if form.is_valid():
                    id_persona = int(request.POST['responsable'])
                    activos = request.POST.getlist('id_activo')
                    productos = request.POST.getlist('producto')
                    # unidadmedida = request.POST.getlist('unidadmedida')
                    # cantidades = request.POST.getlist('cantidad')

                    lista_productos = productos[:-1]
                    count_dict = {}
                    for num in lista_productos:
                        if num in count_dict:
                            count_dict[num] += 1
                        else:
                            count_dict[num] = 1

                    count_dict = {int(key): int(value) for key, value in count_dict.items()}

                    for pro, value in count_dict.items():
                        p = BodegaProducto.objects.get(pk=pro)
                        stock = p.cantidad_total()
                        if value > stock:
                            return JsonResponse({"result": "bad", "mensaje": u"Ha superado la cantidad disponible en bodega de: %s" % p.descripcion})

                    asunto = form.cleaned_data['asunto'] if 'asunto' in request.POST else ''
                    resolucion = form.cleaned_data['resolucion'] if 'resolucion' in request.POST else ''
                    if HdDirector.objects.filter(vigente=True).exists():
                        hdincidente = HdIncidente(asunto=asunto,

                                                persona_id=id_persona,
                                                fechareporte=form.cleaned_data['fechareporte'],
                                                # fechareporte=datetime.now().date(),
                                                horareporte=form.cleaned_data['horareporte'],
                                                medioreporte_id=3,
                                                estado_id=3,
                                                tipoincidente_id=2,
                                                director_id=HdDirector.objects.get(vigente=True,
                                                                                   status=True).persona.id,
                                                revisionequipoexterno=form.cleaned_data[
                                                    'revisionequipoexterno'] if 'revisionequipoexterno' in request.POST else False,
                                                revisionequiposincodigo=form.cleaned_data[
                                                    'revisionequiposincodigo'] if 'revisionequiposincodigo' in request.POST else False,
                                                serie=form.cleaned_data['serie'] if 'serie' in request.POST else '',
                                                )
                        hdincidente.save(request)
                        # incidente.email_notificacion_tic(request.session['nombresistema'])
                        log(u'Adiciono un Detalle Producto Incidente: %s' % hdincidente, request, "add")
                        para = Persona.objects.get(pk=int(request.POST['responsable']))
                        notificacion(
                            asunto,
                            resolucion,
                            para,
                            None,
                            '/mis_activos?action=viewnotificacionlicencia',
                            hdincidente.id,
                            2,
                            'sga',
                            HdIncidente,
                            request
                        )

                        subscriptions = para.usuario.webpush_info.select_related("subscription")
                        push_infos = SubscriptionInfomation.objects.filter(
                            subscription_id__in=subscriptions.values_list('subscription__id', flat=True), app=1,
                            status=True).select_related("subscription")
                        for device in push_infos:
                            payload = {
                                "head": asunto,
                                "body": resolucion,
                                "url": "/mis_activos?action=viewnotificacionlicencia",
                                "action": "loadNotifications",
                            }
                            try:
                                _send_notification(device.subscription, json.dumps(payload), ttl=700)
                            except Exception as exep:
                                print(f"Fallo de envio del push notification: {exep.__str__()}")

                        grupo = HdGrupo.objects.get(pk=1)
                        if HdDetalle_Grupo.objects.filter(responsable=True, grupo=grupo.pk).exists():
                            detalle = HdDetalle_Incidente(incidente_id=hdincidente.id,
                                                          grupo_id=grupo.id,
                                                          agente_id=HdDetalle_Grupo.objects.get(status=True, persona_id=persona.id).id,
                                                          estadoasignacion=1,
                                                          resolucion=form.cleaned_data['resolucion'] if 'resolucion' in request.POST else None,
                                                          # fecharesolucion=form.cleaned_data['fecharesolucion'],
                                                          fecharesolucion=form.cleaned_data['fechareporte'],
                                                          horaresolucion=form.cleaned_data['horareporte'],
                                                          estadoproceso_id= None,
                                                          # responsable_id=HdDetalle_Grupo.objects.get(responsable=True,
                                                          #                                            grupo=grupo.pk).persona.id,
                                                          responsable_id=persona.id,
                                                          estado_id=3
                                                          )
                            detalle.save(request)

                        contador =0
                        for producto in productos:
                            if contador < len(productos)-1:
                                p = BodegaProducto.objects.get(pk=int(producto))

                                # um = BodegaProductoDetalle.objects.get(pk=int(unidadmedida[contador]))
                                # valor_um = um.valor
                                incidente=HdIncidenteProductoDetalle.objects.filter(status=True,activo_id=activos[contador],producto_id=int(producto)).first()
                                if incidente:
                                    raise NameError(f'Licencia de activo y producto seleccionado ya existe | {incidente.activo} : {p}')
                                detproincidente = HdIncidenteProductoDetalle(activo_id=activos[contador],
                                                                             incidente=hdincidente,
                                                                             producto_id=int(producto),
                                                                             unidadmedida_id = 1,
                                                                             cantidad = 1,
                                                                             )
                                detproincidente.save(request)
                                log(u'Adiciono un Detalle Producto Incidente: %s' % detproincidente, request, "add")

                                tipo = 2
                                # cantidad = int(cantidades[contador])
                                # if cantidad > 0:
                                #     cantidad = cantidad * valor_um
                                kardex = actualizar_kardex_e_s(tipo, 1, detproincidente)
                                kardex.save(request)
                                log(u'Adicionar Bodega Kardex: %s' % kardex, request, "add")
                                contador+=1
                            else:
                                break
                        return JsonResponse({"result": "ok"})
                    else:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"No ha definido un Director,registrelo en configuración."})


            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": f"{ex}"})


        elif action == 'edit':
            try:
                incidente = HdIncidente.objects.get(pk=int(request.POST['id']))
                form = HdIncidenciaFrom(request.POST)
                if form.is_valid():
                    activo = None
                    listaayudantes = HdDetalle_Grupo.objects.filter(pk__in=form.cleaned_data['ayudantes'])
                    revisionequipoexterno = form.cleaned_data['revisionequipoexterno'] if 'revisionequipoexterno' in request.POST else False
                    revisionequiposincodigo = form.cleaned_data['revisionequiposincodigo'] if 'revisionequiposincodigo' in request.POST else False
                    if not (revisionequipoexterno or revisionequiposincodigo):
                        if len(request.POST['activo']) > 0:
                            if int(request.POST['activo']) > 0:
                                activo = int(request.POST['activo'])
                    if incidente.tipoincidente_id ==3:
                        # if incidente.h_d_incidentes_set.filter(status=True).exists() and incidente.esta_abierto():
                        if incidente.h_d_incidentes.filter(status=True).exists():
                            detalle = incidente.ultimo_registro()
                            if 'agente' in request.POST:
                                detalle.agente_id = request.POST['agente']
                            if 'grupo' in request.POST:
                                detalle.grupo_id = int(request.POST['grupo'])
                        else:
                            detalle = HdDetalle_Incidente(incidente=incidente,
                                                          agente_id=int(request.POST['agente']),
                                                          grupo_id=int(request.POST['grupo']),
                                                          estadoasignacion=1,
                                                          responsable_id=HdDetalle_Grupo.objects.get(grupo =int(request.POST['grupo']),responsable=True).persona.id)
                            detalle.save(request)
                            if listaayudantes:
                                for ayudante in listaayudantes:
                                    detalleayudantes = HdDetalle_Incidente_Ayudantes(detallleincidente=detalle, agente=ayudante)
                                    detalleayudantes.save(request)
                    # log(u'Modificó incidente: %s' % detalle, request, "edit")
                    incidente.asunto = form.cleaned_data['asunto']
                    incidente.descripcion = form.cleaned_data['descripcion']
                    incidente.fechareporte = form.cleaned_data['fechareporte']
                    incidente.horareporte = form.cleaned_data['horareporte']
                    incidente.subcategoria_id = int(request.POST['subcategoria']) if request.POST['subcategoria'] else None
                    incidente.detallesubcategoria_id = int(request.POST['detallesubcategoria']) if request.POST['detallesubcategoria'] else None
                    incidente.activo_id = activo
                    incidente.medioreporte_id = int(request.POST['medioreporte'])
                    incidente.ubicacion_id = int(request.POST['ubicacion'])
                    incidente.revisionequipoexterno = revisionequipoexterno
                    incidente.revisionequiposincodigo = revisionequiposincodigo
                    incidente.serie = form.cleaned_data['serie'] if 'serie' in request.POST else ''
                    incidente.save(request)
                    log(u'Modificó incidente: %s' % incidente, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'cargar_unidadesmedida':
            try:
                lista = []
                id = int(request.POST['id'])
                p = BodegaProducto.objects.get(pk=id, status=True)
                medidas = BodegaProductoDetalle.objects.filter(producto=p, status=True)
                for m in medidas:
                    lista.append([m.id, m.unidadmedida.descripcion])

                return JsonResponse({'result': 'ok', 'lista': lista})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'cargar_responsable':
            try:
                from sagest.models import ActivoTecnologico
                activo = ActivoTecnologico.objects.get(pk=int(request.POST['id']), status=True)
                per = activo.activotecnologico.responsable
                data = {"result": "ok", "results": {"id": per.id, "name": per.flexbox_repr()}}
                return JsonResponse(data)
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'validar_stock_producto':
            try:
                id = int(request.POST['id'])
                cantidad = int(request.POST['cantidad'])
                valor = BodegaProductoDetalle.objects.get(pk=int(request.POST['unidadmedida'])).valor
                p = BodegaProducto.objects.get(pk=id, status=True)
                cantidadenstock = p.cantidad_total()
                cantidad = cantidadenstock - (cantidad*valor)
                return JsonResponse({'result': 'ok', 'cantidad': cantidad, 'cantidadenstock':cantidadenstock})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})



        elif action == 'reasignaragente':
            try:
                incidente = HdIncidente.objects.get(pk=request.POST['id'])
                form = HdIncidenciaFrom(request.POST)
                if form.is_valid():
                    activo = None
                    listaayudantes = HdDetalle_Grupo.objects.filter(pk__in=form.cleaned_data['ayudantes'])
                    revisionequipoexterno = form.cleaned_data['revisionequipoexterno'] if 'revisionequipoexterno' in request.POST else False
                    revisionequiposincodigo = form.cleaned_data['revisionequiposincodigo'] if 'revisionequiposincodigo' in request.POST else False
                    if not (revisionequipoexterno or revisionequiposincodigo):
                        if not int(request.POST['activo']) > 0 and incidente.tipoincidente_id != 3:
                            return JsonResponse({"result": "bad", "mensaje": u"No ha seleccionado el activo."})
                        if len(request.POST['activo']) > 0:
                            if int(request.POST['activo']) > 0:
                                activo = int(request.POST['activo'])
                    if incidente.tipoincidente:
                        agente = HdDetalle_Grupo.objects.filter(persona_id=persona.id, grupo__tipoincidente=incidente.tipoincidente, status=True)[0]
                    else:
                        agente = HdDetalle_Grupo.objects.filter(persona_id=persona.id, status=True)[0]
                    if incidente.mi_detalle():
                        detalle = incidente.h_d_incidentes.filter(status=True).order_by('-id')[0]
                        if not detalle.grupo:
                            detalle.grupo = agente.grupo
                        if not detalle.agente:
                            detalle.agente = agente
                        detalle.resolucion = form.cleaned_data['resolucion']
                        detalle.fecharesolucion = datetime.now().date()
                        detalle.horaresolucion = datetime.now().time()
                        detalle.estadoasignacion = 2
                        detalle.estado_id = 2
                        detalle.save(request)
                    else:
                        responsable = HdDetalle_Grupo.objects.filter(status=True, grupo=agente.grupo, responsable=True)[0]
                        detalle = HdDetalle_Incidente(incidente=incidente,
                                                      grupo=agente.grupo,
                                                      agente=agente,
                                                      resolucion=form.cleaned_data['resolucion'],
                                                      fecharesolucion=datetime.now().date(),
                                                      horaresolucion=datetime.now().time(),
                                                      estadoasignacion=2,
                                                      responsable=responsable.persona,
                                                      estado_id=2
                                                      )
                        detalle.save(request)
                    responsable = HdDetalle_Grupo.objects.filter(status=True, grupo_id=int(request.POST['grupo']), responsable=True)[0]
                    nuevodetalle = HdDetalle_Incidente(incidente=incidente,
                                                       grupo_id=int(request.POST['grupo']),
                                                       agente_id=int(request.POST['agente']),
                                                       estadoasignacion=1,
                                                       responsable=responsable.persona,
                                                       estado_id=2
                                                       )
                    nuevodetalle.save(request)
                    if listaayudantes:
                        for ayudante in listaayudantes:
                            detalleayudantes = HdDetalle_Incidente_Ayudantes(detallleincidente=nuevodetalle, agente=ayudante)
                            detalleayudantes.save(request)
                    log(u'Reasignó el agente: %s al Agente: %s' % (detalle.agente, nuevodetalle.agente), request,"Reasig")
                    nuevodetalle.email_notificacion_agente_reasignado(request.session['nombresistema'], detalle)
                    incidente = HdIncidente.objects.get(id=detalle.incidente.id)
                    incidente.estado_id = 2
                    incidente.asunto = form.cleaned_data['asunto']
                    incidente.descripcion = form.cleaned_data['descripcion']
                    # incidente.subcategoria_id = int(request.POST['subcategoria']) if request.POST['subcategoria'] else None
                    # incidente.detallesubcategoria_id = int(request.POST['detallesubcategoria']) if request.POST['detallesubcategoria'] else None
                    # incidente.revisionequipoexterno = revisionequipoexterno
                    # incidente.revisionequiposincodigo = revisionequiposincodigo
                    incidente.serie = form.cleaned_data.get('serie', '')
                    if activo:
                        incidente.activo_id = activo
                    incidente.save(request)
                    log(u'Editó por reasignación la cabezera del incidente: %s' % incidente, request, "edit")

                    if incidente.revisionequiposincodigo and incidente.tipoincidente.id == 2:
                        incidente.email_notificacion_equipo_sin_codigo(request.session['nombresistema'])

                    _notificar_reasignacion_incidente(persona=nuevodetalle.agente.persona, incidente=incidente)
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos. %s" % ex.__str__()})

        elif action == 'escalarincidente':
            try:
                incidente = HdIncidente.objects.get(pk=request.POST['id'])
                if not int(request.POST['activo']) > 0 and incidente.tipoincidente_id != 3:
                    return JsonResponse({"result": "bad", "mensaje": u"No ha seleccionado el activo."})
                form = HdIncidenciaFrom(request.POST)
                if form.is_valid():
                    activo = None
                    if len(request.POST['activo']) > 0:
                        if int(request.POST['activo']) > 0:
                            activo = int(request.POST['activo'])
                    if incidente.tipoincidente:
                        agente = HdDetalle_Grupo.objects.filter(persona_id=persona.id, grupo__tipoincidente=incidente.tipoincidente, status=True)[0]
                    else:
                        agente = HdDetalle_Grupo.objects.filter(persona_id=persona.id, status=True)[0]
                    if incidente.mi_detalle():
                        detalle = incidente.h_d_incidentes.filter(status=True).order_by('-id')[0]
                        if not detalle.estadoasignacion == 3:
                            if not detalle.grupo:
                                detalle.grupo = agente.grupo
                            if not detalle.agente:
                                detalle.agente = agente
                            detalle.resolucion = form.cleaned_data['resolucion']
                            detalle.fecharesolucion = datetime.now().date()
                            detalle.horaresolucion = datetime.now().time()
                            detalle.estadoasignacion = 3
                            detalle.estado_id = 2
                            detalle.save(request)
                        else:
                            responsable = HdDetalle_Grupo.objects.filter(status=True, grupo=agente.grupo, responsable=True)[0]
                            detalle = HdDetalle_Incidente(incidente=incidente,
                                                          grupo=agente.grupo,
                                                          agente=agente,
                                                          resolucion=form.cleaned_data['resolucion'],
                                                          fecharesolucion=datetime.now().date(),
                                                          horaresolucion=datetime.now().time(),
                                                          estadoasignacion=3,
                                                          responsable=responsable.persona,
                                                          estado_id=2
                                                          )
                            detalle.save(request)
                    else:
                        responsable = HdDetalle_Grupo.objects.filter(status=True, grupo=agente.grupo, responsable=True)[
                            0]
                        detalle = HdDetalle_Incidente(incidente=incidente,
                                                      grupo=agente.grupo,
                                                      agente=agente,
                                                      resolucion=form.cleaned_data['resolucion'],
                                                      fecharesolucion=datetime.now().date(),
                                                      horaresolucion=datetime.now().time(),
                                                      estadoasignacion=3,
                                                      responsable=responsable.persona,
                                                      estado_id=2
                                                      )
                        detalle.save(request)
                    # nuevodetalle.email_notificacion_escalar(request.session['nombresistema'], detalle)
                    incidente.tipoincidente_id = int(request.POST['tipoincidente'])
                    incidente.estado_id = 2
                    incidente.activo_id = activo
                    incidente.save(request)
                    log(u'El agente: %s escalo el incidente: %s al grupo: %s' % (detalle.agente, incidente, incidente.tipoincidente), request, "Esc")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'resolverincidente':
            try:
                from sagest.models import ActivoTecnologico
                incidente = HdIncidente.objects.get(pk=int(request.POST['id']), status=True)
                # detalle = HdDetalle_Incidente.objects.get(pk=int(request.POST['id']))
                form = HdIncidenciaFrom(request.POST)
                if form.is_valid():
                    getactivo = None
                    listaayudantes = None
                    revisionequipoexterno = form.cleaned_data['revisionequipoexterno'] if 'revisionequipoexterno' in request.POST else False
                    revisionequiposincodigo = form.cleaned_data['revisionequiposincodigo'] if 'revisionequiposincodigo' in request.POST else False
                    if not (revisionequipoexterno or revisionequiposincodigo):
                        if not int(request.POST['activo']) > 0 and incidente.tipoincidente_id != 3:
                            return JsonResponse({"result": "bad", "mensaje": u"No ha seleccionado el activo."})
                        if len(request.POST['activo']) > 0:
                            if int(request.POST['activo']) > 0:
                                getactivo = ActivoFijo.objects.get(pk=int(request.POST['activo']))
                    if incidente.tipoincidente:
                        agente = HdDetalle_Grupo.objects.get(persona=persona, estado=True, grupo__tipoincidente=incidente.tipoincidente)
                    else:
                        agente = HdDetalle_Grupo.objects.get(persona_id=persona.id, estado=True)

                    if incidente.ultimo_registro():
                        listadetalle = incidente.ultimo_registro().hddetalle_incidente_ayudantes_set.values_list('agente_id', flat=True).filter(status=True)
                        listaayudantes = HdDetalle_Grupo.objects.filter(pk__in=listadetalle, status=True)
                    det = HdDetalle_Incidente(incidente=incidente,
                                              agente=agente,
                                              grupo=agente.grupo,
                                              resolucion=request.POST['resolucion'],
                                              fecharesolucion=datetime.now().date(),
                                              horaresolucion=datetime.now().time(),
                                              estadoasignacion=1,
                                              responsable_id=HdDetalle_Grupo.objects.get(responsable=True,grupo=agente.grupo).persona.id,
                                              estadoproceso_id=request.POST['estadobaja'] if 'estadobaja' in request.POST else None,
                                              estado_id=int(request.POST['estado'])
                                              )
                    det.save(request)

                    if listaayudantes:
                        for ayudante in listaayudantes:
                            detalleayudantes = HdDetalle_Incidente_Ayudantes(detallleincidente=det,
                                                                             agente=ayudante
                                                                             )
                            detalleayudantes.save(request)

                    log(u'Resolvio el incidente: %s' % det, request, "resol")
                    if not incidente.tipoincidente:
                        incidente.tipoincidente = agente.grupo.tipoincidente
                    incidente.estado_id = int(request.POST['estado'])
                    if 'subcategoria' in request.POST:
                        if request.POST['subcategoria']:
                            incidente.subcategoria_id = int(request.POST['subcategoria'])
                    if 'detallesubcategoria' in request.POST:
                        if request.POST['detallesubcategoria']:
                            incidente.detallesubcategoria_id = int(request.POST['detallesubcategoria'])
                    if request.POST['causa']:
                        incidente.causa_id = int(request.POST['causa'])
                    else:
                        incidente.causa_id = None
                    incidente.activo_id = getactivo.id if getactivo else getactivo
                    incidente.responsableactivofijo_id = getactivo.responsable.id if getactivo else getactivo
                    incidente.revisionequipoexterno = revisionequipoexterno
                    incidente.revisionequiposincodigo = revisionequiposincodigo
                    incidente.serie = form.cleaned_data['serie'] if 'serie' in request.POST else ''
                    incidente.save(request)
                    # if incidente.ordentrabajo:
                    #      if incidente.estado_id == 3:
                    #         orden = OrdenTrabajo.objects.get(pk=incidente.ordentrabajo.pk)
                    #         orden.estado = 2
                    #         orden.save()
                    # else:
                    #      return JsonResponse({"result": "bad", "mensaje": u"No tiene Orden de trabajo Generada"})
                    log(u'Resolvio el incidente: %s el agente: %s' % (incidente, agente), request, "resol")

                    try:
                        activo_tecnologico = ActivoTecnologico.objects.get(activotecnologico_id=incidente.activo).id
                    except ActivoTecnologico.DoesNotExist:
                        activo_tecnologico = ''

                    productos = request.POST.getlist('producto')
                    unidadmedida = request.POST.getlist('unidadmedida')
                    cantidades = request.POST.getlist('cantidad')
                    contador = 0
                    if int(request.POST['estado'])==3 and len(request.POST.getlist('producto'))>1:
                        for producto in productos:
                            if contador < len(productos) - 1:
                                p = BodegaProducto.objects.get(pk=int(producto))

                                um = BodegaProductoDetalle.objects.get(pk=int(unidadmedida[contador]))
                                valor_um = um.valor
                                detproincidente = HdIncidenteProductoDetalle(incidente=incidente,
                                                                             producto_id=int(producto),
                                                                             unidadmedida_id=int(
                                                                                 unidadmedida[contador]),
                                                                             cantidad=int(cantidades[contador]),
                                                                             )
                                if activo_tecnologico:
                                    detproincidente.activo_id = activo_tecnologico

                                detproincidente.save(request)
                                log(u'Adiciono un Detalle Producto Incidente: %s' % detproincidente, request, "add")

                                tipo = 2
                                cantidad = int(cantidades[contador])
                                if cantidad > 0:
                                    cantidad = cantidad * valor_um
                                kardex = actualizar_kardex_e_s(tipo, cantidad, detproincidente)
                                kardex.save(request)
                                log(u'Adicionar Bodega Kardex: %s' % kardex, request, "add")
                                contador += 1

                    return JsonResponse({"result": "ok"})
                else:
                    return JsonResponse({"result": False, "mensaje": f"Error al validar los datos.", "form": [{k: v[0]} for k, v in f.errors.items()]})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'delincidente':
            try:
                incidente = HdIncidente.objects.get(pk=int(request.POST['id']))
                # if incidente.puede_eliminar():
                log(u'Elimino el incidente: %s' % incidente, request, "del")
                incidente.h_d_incidentes.all().delete()
                incidente.delete()
                if incidente.ordentrabajo:
                    orden = OrdenTrabajo.objects.get(id=incidente.ordentrabajo.id)
                    orden.delete()
                return JsonResponse({"result": "ok", "mensaje": u"Se elimino correctamente.."})
            # else:
            #     return HttpResponse(json.dumps({"result": "bad", "mensaje": u"El Incidente ya a sido topado.."}), content_type="application/json")
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'selectgrupos':
            try:
                if 'id' in request.POST:
                    lista = []
                    listacategoria = []
                    listacausas = []
                    mi_tipo = False
                    grupos = HdGrupo.objects.filter(tipoincidente_id=int(request.POST['id']), status=True)
                    categorias = HdCategoria.objects.filter(tipoincidente_id=int(request.POST['id']), status=True)
                    causas = HdCausas.objects.filter(tipoincidente_id=int(request.POST['id']), status=True)
                    for grupo in grupos:
                        lista.append([grupo.id, grupo.nombre])
                    if causas:
                        for causa in causas:
                            listacausas.append([causa.id, causa.nombre])
                    for categoria in categorias:
                        listacategoria.append([categoria.id, categoria.nombre])
                    if HdDetalle_Grupo.objects.filter(persona=persona, grupo__tipoincidente_id=int(request.POST['id'])).exists():
                        if int(request.POST['id']) == HdDetalle_Grupo.objects.filter(persona=persona,grupo__tipoincidente_id=int(request.POST['id']))[0].grupo.tipoincidente.id:
                            mi_tipo = True
                    return JsonResponse({"result": "ok", 'lista': lista, 'listacausas': listacausas, 'listacategoria': listacategoria, 'mi_tipo': mi_tipo, 'es_tics': True if int(request.POST['id'])== 2 else False})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al consulatar los datos."})

        elif action == 'selectagentes':
            try:
                if 'id' in request.POST:
                    lista = []
                    agentes = HdDetalle_Grupo.objects.filter(grupo_id=int(request.POST['id']), status=True)
                    tiene_responsable = False
                    for agente in agentes:
                        lista.append([agente.id, agente.persona.nombres + ' ' + agente.persona.apellido1 + ' ' + agente.persona.apellido2])
                        if agente.responsable:
                            tiene_responsable = True
                    data = {"results": "ok", 'lista': lista, 'tiene_responsable': tiene_responsable}
                    return JsonResponse(data)
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al consulatar los datos."})

        elif action == 'selectestadoshicidente':
            try:
                lista = []
                estados = HdEstado.objects.filter(status=True).exclude(id__in=[4])
                for estado in estados:
                    lista.append([estado.id, estado.nombre])
                data = {"results": "ok", 'lista': lista}
                return JsonResponse(data)
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al consultar los datos."})

        elif action == 'selectcategoria':
            try:
                if 'id' in request.POST:
                    lista = []
                    subcatecorias = HdSubCategoria.objects.filter(categoria_id=int(request.POST['id']), status=True)
                    for subcat in subcatecorias:
                        lista.append([subcat.id, subcat.nombre])
                    return JsonResponse({"result": "ok", 'lista': lista})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al consulatar los datos."})

        elif action == 'selectdetalle':
            try:
                if 'id' in request.POST:
                    lista = []
                    detalles = HdDetalle_SubCategoria.objects.filter(subcategoria_id=int(request.POST['id']), status=True)
                    for detalle in detalles:
                        lista.append([detalle.id, detalle.nombre])
                    return JsonResponse({"result": "ok", 'lista': lista})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al consulatar los datos."})

        elif action == 'valiprioridad':
            try:
                if 'id' in request.POST:
                    detalle = HdDetalle_SubCategoria.objects.get(pk=int(request.POST['id']), status=True)
                    tiene_prioridad = False
                    pri = ''
                    if detalle.prioridad:
                        tiene_prioridad = True
                        pri = detalle.prioridad.prioridad.nombre
                    data = {"result": "ok", 'tiene_prioridad': tiene_prioridad, 'pri': pri}
                    return JsonResponse(data)
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al consulatar los datos."})

        elif action == 'selectubicacion':
            try:
                if 'id' in request.POST:
                    lista = []
                    ubicaciones = HdBloqueUbicacion.objects.filter(bloque_id=int(request.POST['id']), status=True)
                    for ubi in ubicaciones:
                        lista.append([ubi.id, ubi.ubicacion.nombre])
                    return JsonResponse({"result": "ok", 'lista': lista})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al consulatar los datos."})

        elif action == 'datosactivo':
            try:
                if 'id' in request.POST:
                    activo = ActivoFijo.objects.get(pk=int(request.POST['id']), status=True)
                    fecha = datetime.strftime(activo.fechaingreso, '%Y-%m-%d')
                    vfecha = fecha.split('-')
                    vfecha[0] = str(int(vfecha[0]) + activo.vidautil)
                    fechacaducidad = activo.fecha_caducidad()
                    return JsonResponse(
                        {"result": "ok", 'fechaingreso': str(activo.fechaingreso), 'vidautil': str(activo.vidautil),
                         'tiempo': fechacaducidad})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al consulatar los datos."})

        elif action == 'selectestado':
            try:
                if 'id' in request.POST:
                    lista = []
                    estado = HdEstado.objects.get(pk=int(request.POST['id']), status=True)
                    esta_resuelto = False
                    if estado.id == 3:
                        esta_resuelto = True
                        for lis in HdProceso.objects.filter(status=True):
                            lista.append([lis.id, lis.nombre])
                    return JsonResponse({"result": "ok", 'esta_resuelto': esta_resuelto, "lista": lista})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al consulatar los datos."})

        elif action == 'selectestadoproceso':
            try:
                if 'id' in request.POST:
                    lista = []
                    proceso = HdProceso.objects.get(pk=int(request.POST['id']), status=True)
                    for lis in proceso.hdestado_proceso_set.all():
                        lista.append([lis.id, lis.nombre])
                    return JsonResponse({"result": "ok", "lista": lista})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al consulatar los datos."})

        elif action == 'selectdetalleestpro':
            try:
                if 'id' in request.POST:
                    estado = HdEstado_Proceso.objects.get(pk=int(request.POST['id']), status=True)
                    return JsonResponse({"results": "ok", "detalle": estado.detalle if estado.detalle else ''})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al consulatar los datos."})

        elif action == 'seleccionprioridad':
            try:
                if 'idu' in request.POST and 'idi' in request.POST:
                    prioridad = HdUrgencia_Impacto_Prioridad.objects.get(urgencia_id=int(request.POST['idu']), impacto_id=int(request.POST['idi']))
                    data = {"results": "ok", "tiempo": str(prioridad.prioridad.horamax + ":" + prioridad.prioridad.minutomax + ":" + prioridad.prioridad.segundomax), "prioridad": prioridad.prioridad.nombre}
                    return JsonResponse(data)
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al consulatar los datos."})

        #mantenimientos de estado
        elif action == 'addestado':
            try:
                form = HdEstadoForm(request.POST)
                if form.is_valid():
                    if not HdEstado.objects.filter(status=True, nombre= form.cleaned_data['nombre'].upper()).exists():
                        estado = HdEstado(nombre=form.cleaned_data['nombre'])
                        estado.save(request)
                        if 'imagen' in request.FILES:
                            newfile = request.FILES['imagen']
                            if newfile:
                                newfile._name = generar_nombre("estadohelpdesk_", newfile._name)
                                estado.imagen=newfile
                                estado.save(request)
                                log(u'Agrego una nueva Urgencia: %s' % estado, request, "add")
                            else:
                                return JsonResponse({"result": "bad", "mensaje": u"No a ingresado imagen."})
                        return JsonResponse({"result": "ok"})
                    else:
                        return JsonResponse({"result": "bad", "mensaje": u"El registro ya existe."})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action=='editestado':
            try:
                estado = HdEstado.objects.get(pk=int(request.POST['id']))
                form = HdEstadoEditForm(request.POST)
                if form.is_valid():
                    # if not estado.esta_activo():
                    if not HdEstado.objects.filter(status=True, nombre=form.cleaned_data['nombre'].upper()).exclude(pk=estado.id).exists():
                        estado.nombre = form.cleaned_data['nombre']
                        estado.save(request)
                        log(u'Editar Estado de Help Desk: %s' % estado, request, "edit")
                        return JsonResponse({"result": "ok"})
                    else:
                        return JsonResponse({"result": "bad", "mensaje": u"El nombre del estado está activo."})
                        # else:
                        #     return JsonResponse({"result": "bad", "mensaje": u"El nombre del estado está activo."})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action=='cambiarimagen':
            try:
                estado = HdEstado.objects.get(pk=int(request.POST['id']))
                form = HdEstadoEditForm(request.POST)
                if form.is_valid():
                    if 'imagen' in request.FILES:
                        newfile = request.FILES['imagen']
                        if newfile:
                            newfile._name = generar_nombre("estadohelpdesk_", newfile._name)
                            estado.imagen = newfile
                            estado.save(request)
                            log(u'Editar Estado de Help Desk: %s' % estado, request, "edit")
                            return JsonResponse({"result": "ok"})

                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action=='delestado':
            try:
                estado = HdEstado.objects.get(pk=int(request.POST['id']))
                if not estado.esta_activo():
                    estado.delete()
                    log(u'Elimino estado de Help Desk: %s' % estado, request, "del")
                    return JsonResponse({"result": "ok", "mensaje": u"Se elimino correctamente.."})
                else:
                    return JsonResponse({"result": "bad", "mensaje": u"El estado está activo no se puede eliminar.."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'addmaterialincidente':
            try:
                materialincidente = HdMaterial_Incidente(incidente_id=request.POST['idincidente'],
                                                         material_id=request.POST['idmaterial'],
                                                         unidadmedida_id=request.POST['idunidadmedida'],
                                                         cantidad=request.POST['idcantidad'])
                materialincidente.save(request)
                log(u'Adiciono nuevo materialincidente: %s' % materialincidente, request, "add")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": "Error al guardar los datos."})

        elif action == 'vermaterialesincidente':
            try:
                lista =[]
                incidente = HdIncidente.objects.get(pk=int(request.POST['idincidente']))
                data['materialesincidentes'] = incidente.hdmaterial_incidente_set.filter(status=True)
                template = get_template("helpdesk_hdincidente/consultalistamaterialincidente.html")
                json_content = template.render(data)
                return JsonResponse(
                    {"result": "ok", 'html': json_content, 'title': u'Seleccionar el periodo academico'})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al generar excel."})

        elif action == 'conmaterialesincidentes':
            try:
                materialincidente = HdMaterial_Incidente.objects.get(pk=request.POST['idmater'], status=True)
                nombre = materialincidente.material.nombre
                return JsonResponse({"result": "ok","nombre": nombre })
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": "Error al eliminar los datos."})

        elif action == 'delmaterialincidente':
            try:
                if HdMaterial_Incidente.objects.filter(pk=request.POST['idmaterial'], status=True):
                    materialincidente = HdMaterial_Incidente.objects.get(pk=request.POST['idmaterial'], status=True)
                    materialincidente.delete()
                    log(u'Elimino material incidente: %s' % materialincidente, request, "del")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": "Error al eliminar los datos."})

        # mantenimiento de impacto
        elif action == 'addimpacto':
            try:
                form = HdImpactoForm(request.POST)
                if form.is_valid():
                    if not HdImpacto.objects.filter(status=True, nombre= form.cleaned_data['nombre'].upper()).exists():
                        impacto = HdImpacto(nombre=form.cleaned_data['nombre'],
                                            descripcion=form.cleaned_data['descripcion'],
                                            codigo=form.cleaned_data['codigo'])
                        impacto.save(request)
                        log(u'Agrego una nueva Impacto: %s' % impacto, request, "add")
                        return JsonResponse({"result": "ok"})
                    else:
                        return JsonResponse({"result": "bad", "mensaje": u"El registro ya existe."})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action=='editimpacto':
            try:
                impacto = HdImpacto.objects.get(pk=int(request.POST['id']))
                form = HdImpactoForm(request.POST)
                if form.is_valid():
                    if not impacto.esta_activo():
                        if not HdImpacto.objects.filter(nombre=form.cleaned_data['nombre'], status=True).exclude(pk=impacto.id).exists():
                            impacto.nombre = form.cleaned_data['nombre']
                            impacto.descripcion = form.cleaned_data['descripcion']
                            impacto.codigo = form.cleaned_data['codigo']
                            impacto.save(request)
                            log(u'Editar Impacto: %s' % impacto, request, "editar")
                            return JsonResponse({"result": "ok"})
                        else:
                            return JsonResponse({"result": "bad", "mensaje": u"El Registro ya existe."})
                    else:
                        impacto.descripcion = form.cleaned_data['descripcion']
                        impacto.save(request)
                        log(u'Editar Impacto: %s' % impacto, request, "editar")
                        return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action=='delimpacto':
            try:
                impacto = HdImpacto.objects.get(pk=int(request.POST['id']))
                if not impacto.esta_activo():
                    impacto.delete()
                    log(u'Elimino Impacto: %s' % impacto, request, "del")
                    return JsonResponse({"result": "ok"})
                else:
                    return JsonResponse({"result": "bad", "mensaje": u"No se puede eliminar, el registro ese encuentra activo.."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        #mantenimiento de urgencia

        elif action == 'addurgencia':
            try:
                form = HdUrgenciaForm(request.POST)
                if form.is_valid():
                    if not HdUrgencia.objects.filter(status=True, nombre= form.cleaned_data['nombre'].upper()).exists():
                        urgencia = HdUrgencia(nombre=form.cleaned_data['nombre'],
                                              descripcion=form.cleaned_data['descripcion'],
                                              codigo=form.cleaned_data['codigo']
                                              )
                        urgencia.save(request)
                        log(u'Agrego una nueva Urgencia: %s' % urgencia, request, "add")
                        return JsonResponse({"result": "ok"})
                    else:
                        return JsonResponse({"result": "bad", "mensaje": u"El registro ya existe."})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'addpiezaparte':
            try:
                form = HdPiezaPartesForm(request.POST, request.FILES)
                if 'archivo' in request.FILES:
                    d = request.FILES['archivo']
                    newfilesd = d._name
                    ext = newfilesd[newfilesd.rfind("."):]
                    if ext == '.jpg' or  ext == '.jpeg' or ext == '.png':
                        a = 1
                    else:
                        return JsonResponse({"result": "bad", "mensaje": u"Error, solo archivos .jpg, .jpeg, .png"})
                    if d.size > 2194304:
                        return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 2 Mb."})
                if form.is_valid():
                    if not HdPiezaPartes.objects.filter(status=True, descripcion= form.cleaned_data['descripcion']).exists():
                        piezaparte = HdPiezaPartes(descripcion=form.cleaned_data['descripcion'], estado=form.cleaned_data['estado'])
                        piezaparte.save(request)
                        if 'archivo' in request.FILES:
                            newfile = request.FILES['archivo']
                            newfile._name = generar_nombre("piezapartes_", newfile._name)
                            piezaparte.imagen = newfile
                            piezaparte.save(request)
                        log(u'Agrego una nueva Pieza y Parte: %s' % piezaparte, request, "add")
                        return JsonResponse({"result": "ok"})
                    else:
                        return JsonResponse({"result": "bad", "mensaje": u"El registro ya existe."})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'addcausa':
            try:
                form = HdCausasForm(request.POST)
                if form.is_valid():
                    if not HdCausas.objects.filter(status=True,tipoincidente=form.cleaned_data['tipoincidente'], nombre=form.cleaned_data['nombre']).exists():
                        causa = HdCausas(nombre=form.cleaned_data['nombre'],
                                         tipoincidente=form.cleaned_data['tipoincidente'])
                        causa.save(request)
                        log(u'Agrego una nueva Causa: %s' % causa, request, "add")
                        return JsonResponse({"result": "ok"})
                    else:
                        return JsonResponse({"result": "bad", "mensaje": u"El registro ya existe."})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action=='editcausa':
            try:
                causa = HdCausas.objects.get(pk=int(request.POST['id']))
                form = HdCausasForm(request.POST)
                if form.is_valid():
                    if not HdCausas.objects.filter(status=True,tipoincidente=form.cleaned_data['tipoincidente'], nombre=form.cleaned_data['nombre'].upper()).exclude(pk=causa.id).exists():
                        causa.nombre = form.cleaned_data['nombre']
                        causa.tipoincidente = form.cleaned_data['tipoincidente']
                        causa.save(request)
                        log(u'Editar causa: %s' % causa, request, "edit")
                        return JsonResponse({"result": "ok"})
                    else:
                        return JsonResponse({"result": "bad", "mensaje": u"El nombre de causa ya existe."})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action=='delcausa':
            try:
                causa = HdCausas.objects.get(pk=int(request.POST['id']))
                causa.delete()
                log(u'Elimino causa: %s' % causa, request, "del")
                return JsonResponse({"result": "ok", "mensaje": u"Se elimino correctamente.."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'addfechacierre':
            try:
                form = HdFechacierreForm(request.POST)
                if form.is_valid():
                    fechacierre = HdFechacierresolicitudes(observacion=form.cleaned_data['observacion'],
                                                           fechainicio=form.cleaned_data['fechainicio'],
                                                           activo=form.cleaned_data['activo'],
                                                           fechafin=form.cleaned_data['fechafin'])
                    fechacierre.save(request)
                    log(u'Agrego una nueva fecha de cierre: %s' % fechacierre, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action=='editfechacierre':
            try:
                fechacierre = HdFechacierresolicitudes.objects.get(pk=int(request.POST['id']))
                form = HdFechacierreForm(request.POST)
                if form.is_valid():
                    fechacierre.observacion = form.cleaned_data['observacion']
                    fechacierre.fechainicio = form.cleaned_data['fechainicio']
                    fechacierre.fechafin = form.cleaned_data['fechafin']
                    fechacierre.activo = form.cleaned_data['activo']
                    fechacierre.save(request)
                    log(u'Editar fecha cierre: %s' % fechacierre, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action=='delfechacierre':
            try:
                fechacierre = HdFechacierresolicitudes.objects.get(pk=int(request.POST['id']))
                fechacierre.delete()
                log(u'Elimino fecha cierre: %s' % fechacierre, request, "del")
                return JsonResponse({"result": "ok", "mensaje": u"Se elimino correctamente.."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action=='cerrarfechacierre':
            try:
                fechacierre = HdFechacierresolicitudes.objects.get(pk=int(request.POST['id']))
                fechacierre.estado = False
                fechacierre.activo = False
                fechacierre.usuariocierre = persona.usuario
                fechacierre.fecha_cierre = datetime.now()
                fechacierre.save()
                log(u'Cerro fecha cierre: %s' % fechacierre, request, "del")
                return JsonResponse({"result": "ok", "mensaje": u"Se cerró correctamente.."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'addprecio':
            try:
                form = HdPreciosForm(request.POST)
                if form.is_valid():
                    HdPrecioSolicitudesPiezaPartes.objects.filter(status=True).update(activo=False)
                    if not HdPrecioSolicitudesPiezaPartes.objects.filter(solicitudes_id=int(request.POST['id']),valor=form.cleaned_data['valor'],status=True,).exists():
                        precio = HdPrecioSolicitudesPiezaPartes(solicitudes_id=int(request.POST['id']),
                                                                valor=form.cleaned_data['valor'],
                                                                cierresolicitudes=form.cleaned_data['cierresolicitudes'],
                                                                activo=form.cleaned_data['activo'])
                        precio.save(request)
                        log(u'Agrego un nuevo precio de pieza y parte: %s' % precio.solicitudes, request, "add")
                        return JsonResponse({"result": "ok"})
                    else:
                        return JsonResponse({"result": "bad", "mensaje": u"El registro ya existe."})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action=='editprecio':
            try:
                precio = HdPrecioSolicitudesPiezaPartes.objects.get(pk=int(request.POST['id']))
                form = HdPreciosForm(request.POST)
                if form.is_valid():
                    if form.cleaned_data['activo']:
                        HdPrecioSolicitudesPiezaPartes.objects.filter(status=True).update(activo=False)
                    precio.valor = form.cleaned_data['valor']
                    precio.activo = form.cleaned_data['activo']
                    precio.cierresolicitudes = form.cleaned_data['cierresolicitudes']
                    precio.save(request)
                    log(u'Edito precio: %s' % precio, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action=='delprecio':
            try:
                precio = HdPrecioSolicitudesPiezaPartes.objects.get(pk=int(request.POST['id']))
                precio.delete()
                log(u'Elimino causa: %s' % precio, request, "del")
                return JsonResponse({"result": "ok", "mensaje": u"Se elimino correctamente.."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'addsolicitudpiezaparte':
            try:
                form = HdSolicitudPiezaPartesForm(request.POST)
                if form.is_valid():
                    solicitudpiezaparte = HdSolicitudesPiezaPartes(piezaparte=form.cleaned_data['piezaparte'],
                                                                   grupocategoria=form.cleaned_data['grupocategoria'],
                                                                   capacidad=form.cleaned_data['capacidad'],
                                                                   velocidad=form.cleaned_data['velocidad'],
                                                                   descripcion=form.cleaned_data['descripcion'],
                                                                   tipo=form.cleaned_data['tipo'])
                    solicitudpiezaparte.save(request)
                    log(u'Agrego una nueva solicitud Pieza y Parte: %s' % solicitudpiezaparte, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'addmaterial':
            try:
                form = HdMaterialForm(request.POST)
                if form.is_valid():
                    if not HdMateriales.objects.filter(status=True,nombre=form.cleaned_data['nombre']).exists():
                        material = HdMateriales(nombre=form.cleaned_data['nombre'],
                                                codigo=form.cleaned_data['codigo'])
                        material.save(request)
                        log(u'Agrego una nuevo material: %s' % material, request, "add")
                        return JsonResponse({"result": "ok"})
                    else:
                        return JsonResponse({"result": "bad", "mensaje": u"El registro ya existe."})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action=='editmaterial':
            try:
                material = HdMateriales.objects.get(pk=int(request.POST['id']))
                form = HdMaterialForm(request.POST)
                if form.is_valid():
                    material.codigo = form.cleaned_data['codigo']
                    material.nombre = form.cleaned_data['nombre']
                    material.save(request)
                    log(u'Edito material: %s' % material, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action=='delmaterial':
            try:
                material = HdMateriales.objects.get(pk=int(request.POST['id']))
                material.delete()
                log(u'Elimino material: %s' % material, request, "del")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'editpiezapartes':
            try:
                piezaparte = HdPiezaPartes.objects.get(pk=int(request.POST['id']))
                if 'archivo' in request.FILES:
                    d = request.FILES['archivo']
                    if d.size > 2194304:
                        return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 2 Mb."})
                    else:
                        newfiles = request.FILES['archivo']
                        newfilesd = newfiles._name
                        ext = newfilesd[newfilesd.rfind("."):]
                        if not(ext == '.jpg' or ext == '.jpeg' or ext == '.png'):
                            return JsonResponse({"result": "bad", "mensaje": u"Error, solo archivos .jpg, .jpeg, .png"})
                form = HdPiezaPartesForm(request.POST)
                if form.is_valid():
                    if not HdPiezaPartes.objects.filter(descripcion= form.cleaned_data['descripcion'], status=True).exclude(pk=piezaparte.id).exists():
                        piezaparte.estado = form.cleaned_data['estado']
                        piezaparte.descripcion = form.cleaned_data['descripcion']
                        if 'archivo' in request.FILES:
                            newfile = request.FILES['archivo']
                            newfile._name = generar_nombre("piezapartes_", newfile._name)
                            piezaparte.imagen = newfile
                        piezaparte.save(request)
                        log(u'Editar pieza y partes: %s' % piezaparte, request, "editar")
                        return JsonResponse({"result": "ok"})
                    else:
                        return JsonResponse({"result": "bad", "mensaje": u"El registro ya existe."})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'editsolicitudespiezapartes':
            try:
                solicitudpiezaparte = HdSolicitudesPiezaPartes.objects.get(pk=int(request.POST['id']))
                form = HdSolicitudPiezaPartesForm(request.POST)
                if form.is_valid():
                    solicitudpiezaparte.piezaparte = form.cleaned_data['piezaparte']
                    solicitudpiezaparte.grupocategoria = form.cleaned_data['grupocategoria']
                    solicitudpiezaparte.tipo = form.cleaned_data['tipo']
                    solicitudpiezaparte.capacidad = form.cleaned_data['capacidad']
                    solicitudpiezaparte.velocidad = form.cleaned_data['velocidad']
                    solicitudpiezaparte.descripcion = form.cleaned_data['descripcion']
                    solicitudpiezaparte.save(request)
                    log(u'Editar solicitud pieza y partes: %s' % solicitudpiezaparte, request, "editar")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action=='editurgencia':
            try:
                urgencia = HdUrgencia.objects.get(pk=int(request.POST['id']))
                form = HdUrgenciaForm(request.POST)
                if form.is_valid():
                    if not urgencia.esta_activo():
                        if not HdUrgencia.objects.filter(nombre= form.cleaned_data['nombre'], status=True).exclude(pk=urgencia.id).exists():
                            urgencia.nombre = form.cleaned_data['nombre']
                            urgencia.descripcion = form.cleaned_data['descripcion']
                            urgencia.codigo = form.cleaned_data['codigo']
                            urgencia.save(request)
                            log(u'Editar urgencia: %s' % urgencia, request, "editar")
                            return JsonResponse({"result": "ok"})
                        else:
                            return JsonResponse({"result": "bad", "mensaje": u"El registro ya existe."})
                    else:
                        urgencia.descripcion = form.cleaned_data['descripcion']
                        urgencia.save(request)
                        log(u'Editar urgencia: %s' % urgencia, request, "editar")
                        return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action=='delurgencia':
            try:
                urgencia = HdUrgencia.objects.get(pk=int(request.POST['id']))
                if not urgencia.esta_activo():
                    urgencia.delete()
                    log(u'Elimino Urgencia: %s' % urgencia, request, "del")
                    return JsonResponse({"result": "ok", "mensaje": u"Se elimino correctamente.."})
                else:
                    return JsonResponse({"result": "bad", "mensaje": u"No se puede elimiar el reguistro está activo."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action=='delpiezaparte':
            try:
                piezaparte = HdPiezaPartes.objects.get(pk=int(request.POST['id']))
                piezaparte.delete()
                log(u'Elimino Pieza y Parte: %s' % piezaparte, request, "del")
                return JsonResponse({"result": "ok", "mensaje": u"Se elimino correctamente.."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action=='delsolicitudpiezaparte':
            try:
                solicitudpiezaparte = HdSolicitudesPiezaPartes.objects.get(pk=int(request.POST['id']))
                solicitudpiezaparte.delete()
                log(u'Elimino Solicitud Pieza y Parte: %s' % solicitudpiezaparte, request, "del")
                return JsonResponse({"result": "ok", "mensaje": u"Se elimino correctamente.."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        #mantenimiento de prioridad

        elif action == 'addprioridad':
            try:
                form = HdPrioridadForm(request.POST, request.FILES)
                if form.is_valid():
                    if not HdPrioridad.objects.filter(status=True, nombre= form.cleaned_data['nombre'].upper()).exists():
                        if 'imagen' in request.FILES:
                            newfile = request.FILES['imagen']
                            if newfile:
                                newfile._name = generar_nombre("estadohelpdesk_", newfile._name)
                                prioridad = HdPrioridad(nombre=form.cleaned_data['nombre'],
                                                        codigo=form.cleaned_data['codigo'],
                                                        horamax=form.cleaned_data['hora'] if len(form.cleaned_data['hora']) > 1 else '0' + form.cleaned_data['hora'],
                                                        minutomax=form.cleaned_data['minuto'] if len(form.cleaned_data['minuto']) > 1 else '0' + form.cleaned_data['minuto'],
                                                        segundomax=form.cleaned_data['segundo'] if len(form.cleaned_data['segundo']) > 1 else '0' + form.cleaned_data['segundo'],
                                                        imagen=newfile
                                                        )
                                prioridad.save(request)
                                log(u'Agrego una nueva Prioridad: %s' % prioridad, request, "add")
                            else:
                                return JsonResponse({"result": "bad", "mensaje": u"No a ingresado imagen."})
                        return JsonResponse({"result": "ok"})
                    else:
                        return JsonResponse({"result": "bad", "mensaje": u"El registro ya existe."})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action=='editprioridad':
            try:
                prioridad = HdPrioridad.objects.get(pk=int(request.POST['id']))
                form = HdPrioridadForm(request.POST)
                if form.is_valid():
                    if not prioridad.esta_activo():
                        prioridad.nombre = form.cleaned_data['nombre']
                        prioridad.codigo = form.cleaned_data['codigo']
                        prioridad.horamax = form.cleaned_data['hora'] if len(form.cleaned_data['hora'])>1 else '0'+form.cleaned_data['hora']
                        prioridad.minutomax = form.cleaned_data['minuto'] if len(form.cleaned_data['minuto'])>1 else '0'+form.cleaned_data['minuto']
                        prioridad.segundomax = form.cleaned_data['segundo'] if len(form.cleaned_data['segundo'])>1 else '0'+form.cleaned_data['segundo']
                        prioridad.save(request)
                        log(u'Actualizo Prioridad: %s' % prioridad, request, "editar")
                        return JsonResponse({"result": "ok"})
                    else:
                        prioridad.nombre = form.cleaned_data['nombre']
                        prioridad.horamax = form.cleaned_data['hora'] if len(form.cleaned_data['hora'])>1 else '0'+form.cleaned_data['hora']
                        prioridad.minutomax = form.cleaned_data['minuto'] if len(form.cleaned_data['minuto'])>1 else '0'+form.cleaned_data['minuto']
                        prioridad.segundomax = form.cleaned_data['segundo'] if len(form.cleaned_data['segundo'])>1 else '0'+form.cleaned_data['segundo']
                        prioridad.save(request)
                        log(u'Actualizo Prioridad: %s' % prioridad, request, "editar")
                        return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action=='delprioridad':
            try:
                prioridad = HdPrioridad.objects.get(pk=int(request.POST['id']))
                if not prioridad.esta_activo():
                    prioridad.delete()
                    log(u'Elimino Prioridad: %s' % prioridad, request, "del")
                    return JsonResponse({"result": "ok","mensaje": u".Se elimino correctamente.."})
                else:
                    return JsonResponse({"result": "bad", "mensaje": u"No se puede eliminar, el registro ese encuentra activo."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'cambiarimagenprioridad':
            try:
                prioridad = HdPrioridad.objects.get(pk=int(request.POST['id']))
                form = HdPrioridadImagenForm(request.POST)
                if form.is_valid():
                    if 'imagen' in request.FILES:
                        newfile = request.FILES['imagen']
                        if newfile:
                            newfile._name = generar_nombre("Prioridadhelpdesk_", newfile._name)
                            prioridad.imagen = newfile
                            prioridad.save(request)
                            log(u'Editar Estado de Help Desk: %s' % prioridad, request, "edit")
                            return JsonResponse({"result": "ok"})

                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        #mantenimiento de medio de reporte
        elif action == 'addmedioreporte':
            try:
                form = HdMedioReporteForm(request.POST)
                if form.is_valid():
                    if not HdMedioReporte.objects.filter(status=True, nombre= form.cleaned_data['nombre'].upper()).exists():
                        medio = HdMedioReporte(nombre=form.cleaned_data['nombre'], descripcion=form.cleaned_data['descripcion'])
                        medio.save(request)
                        log(u'Agrego un nuev medio de reporte: %s' % medio, request, "add")
                        return JsonResponse({"result": "ok"})
                    else:
                        return JsonResponse({"result": "bad", "mensaje": u"El registro ya existe."})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action=='editmedioreporte':
            try:
                medio = HdMedioReporte.objects.get(pk=int(request.POST['id']))
                form = HdMedioReporteForm(request.POST)
                if form.is_valid():
                    if not medio.esta_activo():
                        if not HdMedioReporte.objects.filter(status=True, nombre=form.cleaned_data['nombre'].upper()).exclude(pk=medio.id).exists():
                            medio.nombre = form.cleaned_data['nombre']
                            medio.descripcion = form.cleaned_data['descripcion']
                            medio.save(request)
                            log(u'Editar medio de reporte de Help Desk: %s' % medio, request, "edit")
                            return JsonResponse({"result": "ok"})
                        else:
                            return JsonResponse({"result": "bad", "mensaje": u"El Registro está activo."})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action=='delmedioreporte':
            try:
                medio = HdMedioReporte.objects.get(pk=int(request.POST['id']))
                if not medio.esta_activo():
                    medio.delete()
                    log(u'Elimino medio de  reporte de Help Desk: %s' % medio, request, "del")
                    return JsonResponse({"result": "ok", "mensaje": u"Se elimino correctamente.."})
                else:
                    return JsonResponse({"result": "bad", "mensaje": u"El Registro está activo no puede eliminar.."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        #mantenimiento union de urgencia impacto prioridad

        elif action == 'addurgencia_impacto_prioridad':
            try:
                form = HdUrgencia_Impacto_PrioridadForm(request.POST)
                if form.is_valid():
                    if form.cleaned_data['modificar']:
                        gestion = HdUrgencia_Impacto_Prioridad(urgencia=form.cleaned_data['urgencia'],
                                                               impacto=form.cleaned_data['impacto'],
                                                               prioridad=form.cleaned_data['prioridad'],
                                                               modificar=form.cleaned_data['modificar'],
                                                               horamax=form.cleaned_data['hora']  if len(form.cleaned_data['hora'])>1 else '0'+form.cleaned_data['hora'],
                                                               minutomax=form.cleaned_data['minuto']  if len(form.cleaned_data['minuto'])>1 else '0'+form.cleaned_data['minuto'],
                                                               segundomax=form.cleaned_data['segundo']  if len(form.cleaned_data['segundo'])>1 else '0'+form.cleaned_data['segundo'])

                    else:
                        gestion = HdUrgencia_Impacto_Prioridad(urgencia=form.cleaned_data['urgencia'],
                                                               impacto=form.cleaned_data['impacto'],
                                                               prioridad=form.cleaned_data['prioridad'],
                                                               modificar=form.cleaned_data['modificar'])
                    gestion.save(request)
                    log(u'Agrego una nueva Prioridad: %s' % gestion, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action=='editurgencia_impacto_prioridad':
            try:
                gestion = HdUrgencia_Impacto_Prioridad.objects.get(pk=int(request.POST['id']))
                form = HdUrgencia_Impacto_PrioridadForm(request.POST)
                if form.is_valid():
                    if gestion.modificar:
                        gestion.urgencia = form.cleaned_data['urgencia']
                        gestion.impacto = form.cleaned_data['impacto']
                        gestion.prioridad = form.cleaned_data['prioridad']
                        gestion.modificar = form.cleaned_data['modificar']
                        gestion.horamax = form.cleaned_data['hora']
                        gestion.minutomax = form.cleaned_data['minuto']
                        gestion.segundomax = form.cleaned_data['segundo']
                        gestion.save(request)
                        log(u'Editar Prioridad: %s' % gestion, request, "editar")
                        return JsonResponse({"result": "ok"})
                    else:
                        gestion.urgencia = form.cleaned_data['urgencia']
                        gestion.impacto = form.cleaned_data['impacto']
                        gestion.prioridad = form.cleaned_data['prioridad']
                        gestion.modificar = form.cleaned_data['modificar']
                        gestion.save(request)
                        log(u'Editar Prioridad: %s' % gestion, request, "editar")
                        return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action=='delurgencia_impacto_prioridad':
            try:
                gestion = HdUrgencia_Impacto_Prioridad.objects.get(pk=int(request.POST['id']))
                if gestion.esta_activo:
                    gestion.delete()
                    log(u'Elimino Prioridad: %s' % gestion, request, "del")
                    return JsonResponse({"result": "ok", "mensaje": u"Se elimino correctamente.."})
                else:
                    return JsonResponse({"result": "bad", "mensaje": u"El Registro está activo no puede eliminar.."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        #grupo

        elif action == 'addgrupo':
            try:
                form = HdGrupoForm(request.POST)
                if form.is_valid():
                    if not HdGrupo.objects.filter(nombre=form.cleaned_data['grupo'].upper(), status=True).exists():
                        listado = request.POST['otrosintegrantes']
                        grupo = HdGrupo(nombre=form.cleaned_data['grupo'],
                                        descripcion=form.cleaned_data['descripcion'],
                                        tipoincidente_id=request.POST['tipoincidente'])
                        grupo.save(request)
                        log(u'Adiciono nuevo grupo: %s' % grupo, request, "add")
                        if not listado =='addgrupo':
                            for integrante in listado.split(','):
                                admin = Administrativo.objects.get(persona_id= integrante, status=True)
                                detalle= HdDetalle_Grupo(grupo_id=grupo.id, persona_id=admin.persona.id)
                                detalle.save(request)
                                log(u'Adiciono nuevo integrante al  grupo: %s' % detalle, request, "add")
                            return JsonResponse({"result": "ok"})
                        else:
                            return JsonResponse({"result": "ok"})
                    else:
                        return JsonResponse({"result": "bad", "mensaje": u"Ya se registra un grupo con ese nombre."})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'editgrupo':
            try:
                grupo = HdGrupo.objects.get(pk=request.POST['id'])
                form = HdGrupoForm(request.POST)
                if form.is_valid():
                    grupo.nombre = form.cleaned_data['grupo']
                    grupo.descripcion = form.cleaned_data['descripcion']
                    grupo.tipoincidente_id = request.POST['tipoincidente']
                    grupo.save(request)
                    log(u'Modificó departamento: %s' % grupo, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'delgrupo':
            try:
                grupo = HdGrupo.objects.get(pk=request.POST['id'])
                if not grupo.esta_activo():
                    grupo.delete()
                    log(u'Eliminó Departamento: %s' % grupo, request, "del")
                    return JsonResponse({"result": "ok"})
                else:
                    return JsonResponse({"result": "bad", "mensaje": u"El registro se encuentra activo."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'addagente':
            try:
                form = HdDetalle_GrupoForm(request.POST)
                if form.is_valid():
                    admin=Administrativo.objects.get(pk=int(request.POST['id']))
                    grupo = HdGrupo.objects.get(pk=int(request.POST['idg']))
                    if not HdDetalle_Grupo.objects.filter(persona=admin.persona,status=True, grupo__status=True).exists():
                        agente = HdDetalle_Grupo(grupo_id=grupo.id,persona_id=admin.persona.id)
                        agente.save(request)
                        log(u'Adiciono nuevo integrante al grupo: %s' % agente, request, "add")
                        return JsonResponse({"result": "ok"})
                    else:
                        return JsonResponse({"result": "bad", "mensaje": u"Ya se encuentra registrado en el grupo."})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'delagente':
            try:
                agente = HdDetalle_Grupo.objects.get(pk=request.POST['id'])
                if not agente.esta_activo():
                    log(u'Eliminó Agente: %s' % agente, request, "del")
                    agente.delete()
                    return JsonResponse({"result": "ok"})
                else:
                    return JsonResponse({"result": "bad", "mensaje": u"El registro ya se encuentra activo."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'addhdencuesta':
            try:
                form = HdCabEncuestasForm(request.POST)
                if form.is_valid():
                    if not HdCabEncuestas.objects.filter(status=True, nombre=form.cleaned_data['nombre'].upper()).exists():
                        encuestas = HdCabEncuestas(tipoincidente_id=int(request.POST['id']),nombre=form.cleaned_data['nombre'], descripcion = form.cleaned_data['descripcion'], activo = form.cleaned_data['activo'])
                        encuestas.save(request)
                        log(u'Agrego una nueva encuesta: %s' % encuestas, request, "add")
                        return JsonResponse({"result": "ok"})
                    else:
                        return JsonResponse({"result": "bad", "mensaje": u"El registro ya existe."})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'edithdencuesta':
            try:
                encuesta = HdCabEncuestas.objects.get(pk=int(request.POST['id']))
                form = HdCabEncuestasForm(request.POST)
                if form.is_valid():
                    if not HdTipoIncidente.objects.filter(status=True,nombre=form.cleaned_data['nombre'].upper()).exclude(id=encuesta.id).exists():
                        encuesta.nombre = form.cleaned_data['nombre']
                        encuesta.descripcion = form.cleaned_data['descripcion']
                        encuesta.activo = form.cleaned_data['activo']
                        encuesta.save(request)
                        log(u'Editó encuesta: %s' % encuesta, request, "edit")
                        return JsonResponse({"result": "ok"})
                    else:
                        return JsonResponse({"result": "bad", "mensaje": u"El registro ya existe."})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'editdetencuesta':
            try:
                detencuesta = HdDetEncuestas.objects.get(pk=int(request.POST['id']))
                form = HdDetEncuestasForm(request.POST)
                if form.is_valid():
                    detencuesta.pregunta = form.cleaned_data['pregunta']
                    detencuesta.tiporespuesta = form.cleaned_data['tiporespuesta']
                    detencuesta.activo = form.cleaned_data['activo']
                    detencuesta.save(request)
                    log(u'Editó detalle encuesta helpdesk: %s' % detencuesta, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'addhddetencuesta':
            try:
                form = HdDetEncuestasForm(request.POST)
                if form.is_valid():
                    detencuesta = HdDetEncuestas(encuesta_id=request.POST['id'],pregunta=form.cleaned_data['pregunta'],tiporespuesta=form.cleaned_data['tiporespuesta'], activo=form.cleaned_data['activo'])
                    detencuesta.save(request)
                    log(u'Agrego un nuevo detalle de pregunta helpdesk: %s' % detencuesta, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})
        #tipo de incidente
        elif action == 'addtipoincidente':
            try:
                form = HdTipoIncidenteForm(request.POST)
                if form.is_valid():
                    if not HdTipoIncidente.objects.filter(status=True, nombre=form.cleaned_data['nombre'].upper()).exists():
                        tipo = HdTipoIncidente(nombre=form.cleaned_data['nombre'], descripcion = form.cleaned_data['descripcion'])
                        tipo.save(request)
                        log(u'Agrego un nuevo tipo de incidente: %s' % tipo, request, "add")
                        return JsonResponse({"result": "ok"})
                    else:
                        return JsonResponse({"result": "bad", "mensaje": u"El registro ya existe."})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'edittipoincidente':
            try:
                tipo = HdTipoIncidente.objects.get(pk=int(request.POST['id']))
                form = HdTipoIncidenteForm(request.POST)
                if form.is_valid():
                    if not tipo.esta_activo():
                        if not HdTipoIncidente.objects.filter(status=True,nombre=form.cleaned_data['nombre'].upper()).exclude(id=tipo.id).exists():
                            tipo.nombre = form.cleaned_data['nombre']
                            tipo.descripcion = form.cleaned_data['descripcion']
                            tipo.save(request)
                            log(u'Editó Tipo de incidente: %s' % tipo, request, "edit")
                            return JsonResponse({"result": "ok"})
                        else:
                            return JsonResponse({"result": "bad", "mensaje": u"El registro ya existe."})
                    else:
                        tipo.descripcion = form.cleaned_data['descripcion']
                        tipo.save(request)
                        log(u'Editó Tipo de incidente: %s' % tipo, request, "edit")
                        return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'deltipoincidente':
            try:
                tipo = HdTipoIncidente.objects.get(pk=int(request.POST['id']))
                if not tipo.esta_activo():
                    tipo.delete()
                    log(u'Eliminó tipo de incidente: %s' % tipo, request, "del")
                    return JsonResponse({"result": "ok"})
                else:
                    return JsonResponse({"result": "bad", "mensaje": u"No se puede eliminar, el registro ese encuentra activo.."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})
        #categoria

        elif action == 'addcategoria':
            try:
                form = HdCatrgoriaForm(request.POST)
                if form.is_valid():
                    if not HdCategoria.objects.filter(status=True, nombre= form.cleaned_data['nombre'].upper(), tipoincidente_id=request.POST['tipoincidente']).exists():
                        categoria = HdCategoria(nombre=form.cleaned_data['nombre'], tipoincidente_id=request.POST['tipoincidente'])
                        categoria.save(request)
                        log(u'Agrego una nueva Categoria: %s' % categoria, request, "add")
                        return JsonResponse({"result": "ok"})
                    else:
                        return JsonResponse({"result": "bad", "mensaje": u"El registro ya existe."})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'editcategoria':
            try:
                categoria = HdCategoria.objects.get(pk=int(request.POST['id']))
                form = HdCatrgoriaForm(request.POST)
                if form.is_valid():
                    if not categoria.esta_activo():
                        if not HdCategoria.objects.filter(status=True, nombre=form.cleaned_data['nombre'].upper()).exists():
                            categoria.nombre = form.cleaned_data['nombre']
                            categoria.tipoincidente_id = request.POST['tipoincidente']
                            categoria.save(request)
                            log(u'Editar Categoria: %s' % categoria, request, "editar")
                        else:
                            return JsonResponse({"result": "bad", "mensaje": u"El registro ya existe."})
                    else:
                        categoria.tipoincidente_id = request.POST['tipoincidente']
                        categoria.save(request)
                        log(u'Editar Categoria: %s' % categoria, request, "editar")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action=='delcategoria':
            try:
                categoria = HdCategoria.objects.get(pk=int(request.POST['id']))
                if not categoria.esta_activo():
                    categoria.delete()
                    log(u'Elimino categoria: %s' % categoria, request, "del")
                    return JsonResponse({"result": "ok"})
                else:
                    return JsonResponse({"result": "bad", "mensaje": u"No se puede eliminar, el registro ese encuentra activo.."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})
        #Subcategoria
        elif action == 'addsubcategoria':
            try:
                form = HdSubCatrgoriaForm(request.POST)
                if form.is_valid():
                    if not HdSubCategoria.objects.filter(status=True, nombre=form.cleaned_data['subcategoria'].upper()).exists():
                        listado = request.POST['otrosdispositivos']
                        categoria = HdCategoria.objects.get(pk=int(request.POST['id']))
                        subcat = HdSubCategoria(nombre=form.cleaned_data['subcategoria'],categoria=categoria)
                        subcat.save(request)
                        log(u'Agrego una nueva Sub  Categorias: %s' % subcat, request, "add")
                        if not listado == 'addsubcategoria':
                            for dispositivo in listado.split(','):
                                detalle = HdDetalle_SubCategoria(nombre=dispositivo, subcategoria=subcat)
                                detalle.save(request)
                                log(u'Agrego una nueva Iten de Sub categoria: %s' % detalle, request, "add")
                        return JsonResponse({"result": "ok"})
                    else:
                        return JsonResponse({"result": "bad", "mensaje": u"El registro ya existe."})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'editsubcategoria':
            try:
                subcat = HdSubCategoria.objects.get(pk=int(request.POST['id']))
                form = HdSubCatrgoriaForm(request.POST)
                if form.is_valid():
                    if not HdSubCategoria.objects.filter(status=True, nombre= form.cleaned_data['subcategoria'].upper()).exclude(pk=subcat.id).exists():
                        subcat.nombre = form.cleaned_data['subcategoria']
                        subcat.save(request)
                        log(u'Editar Sub Categoria: %s' % subcat, request, "editar")
                        return JsonResponse({"result": "ok"})
                    else:
                        return JsonResponse({"result": "bad", "mensaje": u"El registro ya existe."})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'delsubcategoria':
            try:
                subcat = HdSubCategoria.objects.get(pk=int(request.POST['id']))
                if not subcat.esta_activo():
                    subcat.delete()
                    log(u'Elimino Sub Categoria : %s' % subcat, request, "del")
                    return JsonResponse({"result": "ok"})
                else:
                    return JsonResponse({"result": "bad", "mensaje": u"No se puede eliminar, el registro ese encuentra activo.."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'adddetalle':
            try:
                form = HdDetalleSubCategoriaForm(request.POST)
                if form.is_valid():
                    subcat = HdSubCategoria.objects.get(pk=int(request.POST['id']))
                    prioridad = HdUrgencia_Impacto_Prioridad.objects.get(urgencia_id=int(request.POST['urgencia']),impacto_id=int(request.POST['impacto']))
                    detalle = HdDetalle_SubCategoria(nombre=form.cleaned_data['detalle'],subcategoria_id=subcat.id, prioridad_id=prioridad.id)
                    detalle.save(request)
                    log(u'Agrego una nueva Sub  Categorias: %s' % subcat, request, "add")
                    return JsonResponse({"result": "ok"})

                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'aditdatallesubcategoria':
            try:
                form = HdDetalleSubCategoriaForm(request.POST)
                if form.is_valid():
                    prioridad = HdUrgencia_Impacto_Prioridad.objects.get(urgencia_id=int(request.POST['urgencia']),impacto_id=int(request.POST['impacto']))
                    detalle = HdDetalle_SubCategoria.objects.get(pk=int(request.POST['id']))
                    detalle.nombre = form.cleaned_data['detalle']
                    detalle.prioridad = prioridad
                    detalle.save(request)
                    log(u'Agrego una nueva Sub  Detalle de Sub Categoria: %s' % detalle, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'deldetalle':
            try:
                detalle = HdDetalle_SubCategoria.objects.get(pk=int(request.POST['id']))
                if not detalle.esta_activo():
                    detalle.status = False
                    detalle.save(request)
                    log(u'Elimino Sub Categoria : %s' % detalle, request, "del")
                    return JsonResponse({"result": "ok"})
                else:
                    return JsonResponse({"result": "bad", "mensaje": u"No se puede eliminar, el registro ese encuentra activo.."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'addproceso':
            try:
                form = HdProcesoFrom(request.POST)
                if form.is_valid():
                    if not HdProceso.objects.filter(nombre= request.POST['nombre']).exists():
                        proceso = HdProceso(nombre= request.POST['nombre'])
                        proceso.save(request)
                        log(u'Agrego un nuevo proceso de incidente: %s' % proceso, request, "add")
                        return JsonResponse({"result": "ok"})
                    else:
                        return JsonResponse({"result": "bad", "mensaje": u"El El Nombre del proceso ya existe."})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'editproceso':
            try:
                proceso = HdProceso.objects.get(pk=int(request.POST['id']))
                if not proceso.esta_activo():
                    if not HdProceso.objects.filter(nombre=request.POST['nombre']).exclude(pk=int(request.POST['id'])).exists():
                        proceso.nombre=request.POST['nombre']
                        proceso.save(request)
                        log(u'Elimino Proceso Incidente: %s' % proceso, request, "del")
                        return JsonResponse({"result": "ok"})
                    else:
                        return JsonResponse({"result": "bad", "mensaje": u"No se Guardar, El nombre del proceso ya existe.."})
                else:
                    return JsonResponse({"result": "bad", "mensaje": u"No se puede Editar, el registro ese encuentra activo.."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'delproceso':
            try:
                proceso = HdProceso.objects.get(pk=int(request.POST['id']))
                if not proceso.esta_activo():
                    log(u'Elimino Proceso Incidente: %s' % proceso, request, "del")
                    proceso.delete()
                    return JsonResponse({"result": "ok"})
                else:
                    return JsonResponse({"result": "bad", "mensaje": u"No se puede eliminar, el registro ese encuentra activo.."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'addestado_proceso':
            try:
                form = HdEstadoProcesoFrom(request.POST)
                if form.is_valid():
                    if not HdEstado_Proceso.objects.filter(nombre=form.cleaned_data['nombre'].upper(), proceso_id=int(request.POST['id']),status=True).exists():
                        estadopro = HdEstado_Proceso(nombre= request.POST['nombre'], proceso_id= int(request.POST['id']),detalle=request.POST['detalle'])
                        estadopro.save(request)
                        log(u'Agrego un nuevo estado de proceso de incidente: %s' % estadopro, request, "add")
                        return JsonResponse({"result": "ok"})
                    else:
                        return JsonResponse({"result": "bad", "mensaje": u"El registro ya existe.."})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'editestado_proceso':
            try:
                estadopro = HdEstado_Proceso.objects.get(pk=int(request.POST['id']))
                form = HdEstadoProcesoFrom(request.POST)
                if form.is_valid():
                    if not estadopro.esta_activo():
                        if not HdEstado_Proceso.objects.filter(nombre=form.cleaned_data['nombre'].upper(),proceso_id=int(request.POST['id']),status=True).exists():
                            estadopro.nombre=request.POST['nombre']
                            estadopro.detalle=request.POST['detalle']
                            estadopro.save(request)
                            log(u'Editar Estado de proceso: %s' % estadopro, request, "del")
                            return JsonResponse({"result": "ok"})
                        else:
                            return JsonResponse({"result": "bad", "mensaje": u"No se puede guadar, ya existe el reguistro.."})
                    else:
                        return JsonResponse({"result": "bad", "mensaje": u"No se puede guadar, ya se encuentra activo.."})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'delestado_proceso':
            try:
                estadopro = HdEstado_Proceso.objects.get(pk=int(request.POST['id']))
                if not  estadopro.esta_activo():
                    log(u'Elimino Proceso Incidente: %s' % estadopro, request, "del")
                    estadopro.delete()
                    return JsonResponse({"result": "ok"})
                else:
                    return JsonResponse({"result": "bad", "mensaje": u"No se puede eliminar, el registro ese encuentra activo.."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'addagregarestado_proceso':
            try:
                form = HdEstadoProcesoFrom(request.POST)
                if form.is_valid():
                    if not HdEstado_Proceso.objects.filter(nombre=form.cleaned_data['nombre'].upper(),proceso_id=int(request.POST['id']),status=True).exists():
                        estadopro = HdEstado_Proceso(nombre=request.POST['nombre'],proceso_id=int(request.POST['id']))
                        estadopro.save(request)
                        log(u'Agrego un nuevo estado de proceso de incidente: %s' % estadopro, request, "add")
                        return JsonResponse({"result": "ok"})
                    else:
                        return JsonResponse({"result": "bad", "mensaje": u"El registro ya existe.."})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'actualizar_Director':
            try:
                if 'id' in request.POST:
                    res = HdDirector.objects.get(pk=int(request.POST['id']))
                    if HdDirector.objects.filter(status=True,vigente=True).exists():
                        director = HdDirector.objects.filter(status=True, vigente=True)
                        for dir in director:
                            dir.vigente=False
                            dir.save(request)
                    res.vigente=True
                    res.save(request)
                    log(u'Actualizar  Director: %s' % res, request, "edit")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'validar_responsable':
            try:
                res = HdDetalle_Grupo.objects.get(pk=int(request.POST['id']))
                if res.estado:
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al consultar los datos."})

        elif action == 'actualizar_responsable':
            try:
                if 'id' in request.POST:
                    res = HdDetalle_Grupo.objects.get(pk=int(request.POST['id']))
                    if res.estado:
                        grupo = HdGrupo.objects.get(pk=res.grupo.id)
                        if grupo.hddetalle_grupo_set.filter(responsable=True, estado=True).exists():
                            agentes = grupo.hddetalle_grupo_set.filter(responsable=True, estado=True)
                            for agente in agentes:
                                agente.responsable=False
                                agente.save(request)
                        res.responsable=True
                        res.save(request)
                        log(u'Actualizar Responsable: %s' % res, request, "edit")
                        return JsonResponse({"result": "ok"})
                    else:
                        return JsonResponse({"result": "bad", "mensaje": u"No se puede asignar como responsable, porque debe estar activo.."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'actualizar_agente':
            try:
                if 'id' in request.POST:
                    agente = HdDetalle_Grupo.objects.get(pk=int(request.POST['id']))
                    if agente.estado:
                        agente.estado=False
                    else:
                        agente.estado=True
                    agente.save(request)
                    log(u'Actualizo el agente: %s' % agente, request, "edit")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'adddirector':
            try:
                form = HdDirectorForm(request.POST)
                if form.is_valid():
                    admin = Administrativo.objects.get(pk=int(request.POST['id']))
                    director = HdDirector(persona=admin.persona)
                    director.save(request)
                    log(u'Adicionó Director: %s' % director, request, "add")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'editdirector':
            try:
                form = HdDirectorForm(request.POST)
                if form.is_valid():
                    admin = Administrativo.objects.get(pk=int(request.POST['id']))
                    director = HdDirector.objects.get(pk=int(request.POST['idd']))
                    director.persona_id=admin.persona.id
                    director.save(request)
                    log(u'Modificó Director: %s' % director, request, "add")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'deldirector':
            try:
                director = HdDirector.objects.get(pk=int(request.POST['id']))
                log(u'Elimino director: %s' % director, request, "del")
                director.delete()
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'addbloque':
            try:
                if 'bloque' in request.POST:
                    if not HdBloque.objects.filter(nombre=request.POST['bloque'], status=True).exists():
                        bloque = HdBloque(nombre=request.POST['bloque'])
                        bloque.save(request)
                        log(u'Agrego un nuevo estado de proceso de incidente: %s' % bloque, request, "add")
                        return JsonResponse({"result": "ok"})
                    else:
                        return JsonResponse({"result": "bad", "mensaje": u"El registro ya existe.."})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'editbloque':
            try:
                if 'bloque' in request.POST:
                    if not HdBloque.objects.filter(nombre=request.POST['bloque'], status=True).exclude(pk=int(request.POST['id'])).exists():
                        bloque = HdBloque.objects.get(pk=int(request.POST['id']))
                        bloque.nombre=request.POST['bloque']
                        bloque.save(request)
                        log(u'Edito la bloque: %s' % bloque, request, "edit")
                        return JsonResponse({"result": "ok"})
                    else:
                        return JsonResponse({"result": "bad", "mensaje": u"El registro ya existe.."})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'delbloque':
            try:
                blo = HdBloque.objects.get(pk=int(request.POST['id']))
                if blo.puede_eliminar_bloque():
                    return JsonResponse({"result": "bad", "mensaje": u"No puene eliminar, ya esta en uso.."})
                log(u'Elimino bloque: %s' % blo, request, "del")
                blo.delete()
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'addubicacion':
            try:
                if 'ubicacion' in request.POST:
                    if not HdUbicacion.objects.filter(nombre=request.POST['ubicacion'], status=True).exists():
                        ubicacion = HdUbicacion(nombre=request.POST['ubicacion'])
                        ubicacion.save(request)
                        log(u'Agrego un nuevo estado de proceso de incidente: %s' % ubicacion, request, "add")
                        return JsonResponse({"result": "ok"})
                    else:
                        return JsonResponse({"result": "bad", "mensaje": u"El registro ya existe.."})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'editubicacion':
            try:
                if 'ubicacion' in request.POST:
                    if not HdUbicacion.objects.filter(nombre=request.POST['ubicacion'], status=True).exclude(pk=int(request.POST['id'])).exists():
                        ubicacion = HdUbicacion.objects.get(pk=int(request.POST['id']))
                        ubicacion.nombre=request.POST['ubicacion']
                        ubicacion.save(request)
                        log(u'Edito la ubicacion: %s' % ubicacion, request, "edit")
                        return JsonResponse({"result": "ok"})
                    else:
                        return JsonResponse({"result": "bad", "mensaje": u"El registro ya existe.."})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'delubicacion':
            try:
                ubi = HdUbicacion.objects.get(pk=int(request.POST['id']))
                if ubi.puede_eliminar_ubicacion():
                    return JsonResponse({"result": "bad", "mensaje": u"No puene eliminar, ya esta en uso.."})
                log(u'Elimino ubicacion: %s' % ubi, request, "del")
                ubi.delete()
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'addbloque_ubicacion':
            try:
                if 'lista' in request.POST:
                    for x in json.loads(request.POST['lista']):
                        bloque = HdBloqueUbicacion(bloque_id=x['idbloque'], ubicacion_id=x['idubicacion'])
                        bloque.save(request)
                        log(u'Configuro nuevo bloque: %s' % bloque, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'delbloque_ubicacion':
            try:
                bloque = HdBloqueUbicacion.objects.get(pk=int(request.POST['id']))
                log(u'Elimino ubicacion: %s' % bloque.ubicacion, request, "del")
                bloque.delete()
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'delgrupobloque_ubicacion':
            try:
                bloque = HdBloqueUbicacion.objects.get(pk=int(request.POST['id']))
                listabloques = HdBloqueUbicacion.objects.filter(status=True, bloque=bloque.bloque)
                log(u'Elimino ubicacion: %s' % bloque.bloque, request, "del")
                listabloques.delete()
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'pdflistapartes':
            try:
                data = {}
                data['fechaactual'] = datetime.now()
                data['nombrespersona'] = persona
                grupousuario = persona.hdpersona.filter(status=True)[0].grupo
                data['nombrespersonaresponsable'] = HdDetalle_Grupo.objects.filter(grupo=grupousuario,responsable=True,estado=True,status=True)[0].persona
                data['fechainicio'] = datetime.strptime(request.POST['ini'], "%d-%m-%Y").date()
                data['fechafin'] = datetime.strptime(request.POST['fin'], "%d-%m-%Y").date()
                data['totalpiezaspartes'] = requerimientos = HdRequerimientosPiezaPartes.objects.values_list('solicitudes__piezaparte__descripcion','solicitudes__tipo','solicitudes__capacidad','solicitudes__velocidad','preciosolicitud__valor','solicitudes__descripcion').filter(Q (incidente__activo__archivobaja__isnull=True) | Q (incidente__activo__archivobaja=''),preciosolicitud__status=True,incidente__fechareporte__range=(convertir_fecha(request.POST['ini']), convertir_fecha(request.POST['fin'])), incidente__status=True, status=True).annotate(sumuni=Sum('preciosolicitud__valor')).annotate(contador=Count('preciosolicitud__valor'))
                data['totalprecioreferencial'] = requerimientos.values_list('sumuni').aggregate(totalprecio=Sum('sumuni'))['totalprecio']
                data['totalprecioreferencialunitario'] = requerimientos.values_list('preciosolicitud__valor').aggregate(totalsolicitudvalor=Sum('preciosolicitud__valor'))['totalsolicitudvalor']
                data['totalcantidad'] = requerimientos.values_list('contador').aggregate(totalcontador=Sum('contador'))['totalcontador']
                return conviert_html_to_pdf(
                    'helpdesk_hdincidente/pdflistapartes.html',
                    {
                        'pagesize': 'A4',
                        'data': data,
                    }
                )
            except Exception as ex:
                pass

        elif action == 'pdflistaencuesta':
            try:
                fecha = datetime.now().date()
                grupo = HdDetalle_Grupo.objects.get(estado=True, persona=persona).grupo
                mesfecha = MESES_CHOICES[fecha.month - 1][1]
                incidentes = HdRespuestaEncuestas.objects.values_list('cabrespuesta__incidente_id','respuesta__nombre').filter(cabrespuesta__incidente__h_d_incidentes__estado_id=3,cabrespuesta__incidente__h_d_incidentes__agente__grupo=grupo,cabrespuesta__incidente__fechareporte__range=(convertir_fecha(request.POST['ini']), convertir_fecha(request.POST['fin'])), cabrespuesta__incidente__estado=3).distinct().annotate(contado=Count('id')).order_by('respuesta__id')
                reporterespuesta = incidentes.values_list('respuesta__nombre').distinct().annotate(contado=Count('id', distinct=True))
                total = reporterespuesta.aggregate(total=Sum('contado'))['total']
                add_titulo_reportlab(descripcion = "MEMORANDO nº: UNEMI-TIC-"+ str(fecha.year) +"-"+ str(fecha.day) +"-___", tamano = 10, alineacion = 2, espacios = 19,)
                add_titulo_reportlab(descripcion = "Milagro, "+ str(fecha.day) +" de "+ mesfecha +" del "+ str(fecha.year) +".", tamano = 10, alineacion = 2, espacios = 19, afterespacio=15)
                add_titulo_reportlab(descripcion = "ENCUESTA DE NIVEL DE SASTIFACCIÓN AL USUARIO EN RESOLUCIÓN DE INCIDENTES TEGNOLÓGICOS", tamano = 14, espacios = 19, afterespacio=15)
                add_titulo_reportlab(descripcion = "FECHA DE CORTE: " + convertir_fecha(request.POST['ini']).strftime("%d-%m-%Y") + " al " + convertir_fecha(request.POST['fin']).strftime("%d-%m-%Y"), tamano = 14, espacios = 19, afterespacio=15)
                add_tabla_reportlab(encabezado = [('Nº','RESPUESTAS','CANTIDAD','PORCENTAJE')],
                                    detalles = [(contador + 1, agente[0], agente[1], round(((agente[1]*100)/total),2)) for contador, agente in  enumerate(reporterespuesta)],
                                    anchocol = [50, 250, 100,100],
                                    cabecera_left_center = [True, False, True, True],
                                    detalle_left_center = [True, False, True, True])
                add_tabla_reportlab(encabezado=[('', 'TOTAL', total, '100')],
                                    detalles='',
                                    anchocol=[50, 250, 100, 100], cabecera_left_center=[True, True, True, True],
                                    detalle_left_center=[True, False, True, True])
                add_graficos_circular_reporlab(datavalor = [(agente[1]+0) for agente in reporterespuesta],
                                               datanombres = [u'%s'% (agente[0]) for agente in reporterespuesta],
                                               anchografico = 150, altografico = 150,
                                               posiciongrafico_x = 120,  posiciongrafico_y = 20,
                                               titulo = 'RESPUESTAS', tamanotitulo = 10,
                                               ubicaciontitulo_x = 90, ubicaciontitulo_y = 12)
                listadoagentes = HdDetalle_Incidente.objects.values_list('agente__persona__id','agente__persona__apellido1','agente__persona__apellido2','agente__persona__nombres').filter(incidente__in=incidentes.values_list('cabrespuesta__incidente_id')).filter(estado_id=3).distinct()
                for idlista in listadoagentes:
                    detrespuestas = HdRespuestaEncuestas.objects.values_list('respuesta__nombre').filter(cabrespuesta__incidente__h_d_incidentes__estado_id=3,cabrespuesta__incidente__h_d_incidentes__agente__persona_id=idlista[0], cabrespuesta__incidente__in=incidentes.values_list('cabrespuesta__incidente_id'), cabrespuesta__incidente__estado=3).distinct().annotate(contado=Count('id', distinct=True)).order_by('respuesta__id')
                    totaldetalle = detrespuestas.aggregate(total=Sum('contado'))['total']
                    add_titulo_reportlab(descripcion="_____________________________________________________________________________________________________", tamano=9,alineacion=0, espacios=15, afterespacio=5)
                    add_titulo_reportlab(descripcion="ENCUESTA DE NIVEL DE SASTIFACCIÓN AL USUARIO EN RESOLUCIÓN DE INCIDENTES TEGNOLÓGICOS", tamano=12, espacios=19, afterespacio=15)
                    add_titulo_reportlab(descripcion="AGENTE: "+(idlista[1] + ' ' + idlista[2] + ' ' + idlista[3]), tamano=9,alineacion=0, espacios=15, afterespacio=5)
                    add_tabla_reportlab(encabezado=[('Nº', 'RESPUESTAS', 'CANTIDAD','PROCENTAJE')],
                                        detalles=[(contador + 1, subcategoria[0], subcategoria[1], round(((subcategoria[1]*100)/totaldetalle),2)) for contador, subcategoria in enumerate(detrespuestas)],
                                        anchocol=[50, 250, 100,100], cabecera_left_center=[True, False, True, True],
                                        detalle_left_center=[True, False, True, True])
                    add_tabla_reportlab(encabezado=[('', 'TOTAL', totaldetalle, '100')],
                                        detalles='',
                                        anchocol=[50, 250, 100, 100], cabecera_left_center=[True, True, True, True],
                                        detalle_left_center=[True, False, True, True])
                    add_graficos_circular_reporlab(datavalor=[(formatean[1]) for formatean in detrespuestas],
                                                   datanombres=[u'%s' % (formatean[0]) for formatean in
                                                                detrespuestas],
                                                   anchografico=150, altografico=150,
                                                   posiciongrafico_x=120, posiciongrafico_y=20,
                                                   titulo='RESPUESTA ENCUESTA', tamanotitulo=10,
                                                   ubicaciontitulo_x=90, ubicaciontitulo_y=12)
                return generar_pdf_reportlab()
            except Exception as ex:
                return HttpResponseRedirect("/helpdesk_hdincidente?info=%s" % 'Error al generar informe de incidente ')

        elif action == 'informeincidente':
            try:
                grupo = HdDetalle_Grupo.objects.get(estado=True,persona=persona).grupo
                fechainicio = datetime.strptime(request.POST['ini'], "%d-%m-%Y").date()
                fechafin = datetime.strptime(request.POST['fin'], "%d-%m-%Y").date()
                if fechafin >= datetime.now().date():
                    fechafin = datetime.now().date() + timedelta(days=-1)
                listadoinformebaja = InformeActivoBaja.objects.values_list('usuario_creacion__persona__apellido1','usuario_creacion__persona__apellido2','usuario_creacion__persona__nombres').filter(fecha_creacion__range=(fechainicio, fechafin), status=True).annotate(contado=Count('usuario_creacion__id')).order_by('usuario_creacion__persona__apellido1','usuario_creacion__persona__apellido2')
                incidentes = HdIncidente.objects.filter(fechareporte__range=(fechainicio, fechafin), estado=3, h_d_incidentes__estado_id=3,h_d_incidentes__agente__grupo=grupo).distinct()
                agentesatendieron = incidentes.values_list('h_d_incidentes__agente__persona__apellido1', 'h_d_incidentes__agente__persona__apellido2', 'h_d_incidentes__agente__persona__nombres', 'h_d_incidentes__agente__persona_id').filter(h_d_incidentes__estado_id=3).annotate(contado=Count('h_d_incidentes__agente__id')).order_by('h_d_incidentes__agente__persona__apellido1', 'h_d_incidentes__agente__persona__apellido2', 'h_d_incidentes__agente__persona__nombres','contado')
                listadocausas = incidentes.values_list('causa__nombre').filter(h_d_incidentes__estado_id=3,causa__status=True).distinct().annotate(contado=Count('id', distinct=True))
                totalcausas = listadocausas.aggregate(total=Sum('contado'))['total']
                totalagentesatendieron = 0
                promediosuma = 0
                listatendieron = []
                totalagentesatendieron = agentesatendieron.aggregate(total=Sum('contado'))['total']
                # for listadoatendieron in agentesatendieron:
                #     totalagentesatendieron = totalagentesatendieron + listadoatendieron[3]
                idagentesatendieron = agentesatendieron.values_list('h_d_incidentes__agente__persona_id',flat=True)
                agentesgrupo = HdDetalle_Grupo.objects.values_list('persona__apellido1','persona__apellido2','persona__nombres','persona_id').filter(grupo=grupo, responsable=False, estado=True, status=True).exclude(persona__id__in=idagentesatendieron)
                userformatean = incidentes.values_list('h_d_incidentes__agente__persona__apellido1', 'h_d_incidentes__agente__persona__apellido2', 'h_d_incidentes__agente__persona__nombres').filter(h_d_incidentes__estado_id=3,detallesubcategoria__subcategoria_id=31,status=True).annotate(contado=Count('h_d_incidentes__agente__id')).order_by('h_d_incidentes__agente__persona__apellido1', 'h_d_incidentes__agente__persona__apellido2', 'h_d_incidentes__agente__persona__nombres','contado')
                totaluserformatean = userformatean.aggregate(total=Sum('contado'))['total']
                for listadoatendieron in agentesatendieron:
                    listatendi = []
                    promediosuma = round(((listadoatendieron[4] * 100)/totalagentesatendieron),2)
                    for lis in listadoatendieron:
                        listatendi.append(lis)
                    listatendi.append(promediosuma)
                    listatendieron.append((listatendi))
                for agen in agentesgrupo:
                    listatendies = []
                    for agenlis in agen:
                        listatendies.append(agenlis)
                    listatendies.append('0')
                    listatendies.append('0')
                    listatendies.append('0')
                    listatendieron.append((listatendies))
                categoriasatendida = incidentes.values_list('detallesubcategoria__subcategoria__categoria__nombre','detallesubcategoria__subcategoria__categoria__orden').filter(detallesubcategoria__subcategoria__categoria__nombre__isnull=False).annotate(contado=Count('detallesubcategoria__subcategoria__categoria__id')).order_by('detallesubcategoria__subcategoria__categoria__orden','detallesubcategoria__subcategoria__categoria__nombre', 'contado')
                add_titulo_reportlab(descripcion = "INFORME DE INCIDENTES", tamano = 16,  espacios = 19)
                add_titulo_reportlab(descripcion = "GRUPO DE " + grupo.nombre + " " + fechainicio.strftime("%d-%m-%Y") + " al " + fechafin.strftime("%d-%m-%Y"), tamano = 14, espacios = 19)
                add_tabla_reportlab(encabezado = [('Nº','INCIDENTES RESUELTOS POR AGENTE','VALOR','%')],
                                    detalles = [(contador + 1, u'%s %s %s'% (agente[0], agente[1], agente[2]), agente[4], agente[5]) for contador, agente in  enumerate(listatendieron)],
                                    anchocol = [50, 300, 100,50],
                                    cabecera_left_center = [True, False, True, True],
                                    detalle_left_center = [True, False, True, True])
                add_graficos_circular_reporlab(datavalor = [(agente[4]+0) for agente in agentesatendieron],
                                               datanombres = [u'%s %s'% (agente[0], agente[1]) for agente in agentesatendieron],
                                               anchografico = 150, altografico = 150,
                                               posiciongrafico_x = 120,  posiciongrafico_y = 20,
                                               titulo = 'INCIDENTES RESUELTOS POR AGENTE', tamanotitulo = 10,
                                               ubicaciontitulo_x = 90, ubicaciontitulo_y = 12)
                add_titulo_reportlab(descripcion="_____________________________________________________________________________________________________", tamano=9, alineacion=0, espacios=15, afterespacio=20)
                add_tabla_reportlab(encabezado = [('Nº', 'CATEGORIAS ATENDIDAS', 'VALOR')],
                                    detalles = [(contador + 1, categoria[0], categoria[2]) for contador, categoria in enumerate(categoriasatendida)],
                                    anchocol = [50, 300, 100],
                                    cabecera_left_center = [True, False, True],
                                    detalle_left_center = [True, False, True])
                add_graficos_barras_reportlab(datavalor = [[categoria[2] for categoria in categoriasatendida]],
                                              datanombres = [u'%s '% categoria[0] for categoria in categoriasatendida],
                                              anchografico = 300, altografico = 125, tamanoletra = 6,
                                              posiciongrafico_x = 200, posiciongrafico_y = 30,
                                              titulo = 'CATEGORIAS ATENDIDAS', tamanotitulo = 10,
                                              ubicaciontitulo_x = 140, ubicaciontitulo_y = 12, mostrarleyenda=False)


                listacategoria = HdCategoria.objects.filter(tipoincidente_id=2,status=True).order_by('orden')
                for idlista in listacategoria:
                    subcategoriasatendidas = incidentes.values_list('detallesubcategoria__subcategoria__id','detallesubcategoria__subcategoria__nombre').filter(detallesubcategoria__subcategoria__id__isnull=False,detallesubcategoria__subcategoria__categoria__id=idlista.id).annotate(contado=Count('detallesubcategoria__subcategoria__id')).order_by('detallesubcategoria__subcategoria__nombre', 'contado')
                    if subcategoriasatendidas:
                        add_titulo_reportlab(descripcion="_____________________________________________________________________________________________________", tamano=9, alineacion=0, espacios=15, afterespacio=20)
                        add_tabla_reportlab(encabezado=[('Nº', idlista.nombre , 'VALOR')],
                                            detalles=[(contador + 1, subcategoria[1], subcategoria[2]) for
                                                      contador, subcategoria in enumerate(subcategoriasatendidas)],
                                            anchocol=[50, 300, 100], cabecera_left_center=[True, False, True],
                                            detalle_left_center=[True, False, True])
                        add_graficos_barras_reportlab(datavalor=[[subcategoria[2] for subcategoria in subcategoriasatendidas]],
                                                      datanombres=[u'%s' % subcategoria[1] for subcategoria in
                                                                   subcategoriasatendidas],
                                                      anchografico=300, altografico=125, tamanoletra=6,
                                                      posiciongrafico_x=200, posiciongrafico_y=30,
                                                      titulo=idlista.nombre, tamanotitulo=10,
                                                      ubicaciontitulo_x=140, ubicaciontitulo_y=12, posicionleyenda_x=430,
                                                      mostrarleyenda=False)
                add_titulo_reportlab(descripcion="_____________________________________________________________________________________________________",  tamano=9, alineacion=0, espacios=15, afterespacio=20)
                add_tabla_reportlab(encabezado=[('Nº', 'SISTEMA OPERATIVO Y APLICATIVOS', 'TOTAL', '%')],
                                    detalles=[(contador1 + 1, formatean[0] + ' ' + formatean[1] + ' ' + formatean[2], formatean[3],round(((formatean[3] * 100)/totaluserformatean),2)) for contador1, formatean in enumerate(userformatean)],
                                    anchocol=[50, 250, 100, 50],
                                    cabecera_left_center=[True, False, True, True],
                                    detalle_left_center=[True, False, True, True])
                add_graficos_circular_reporlab(datavalor=[(formatean[3]) for formatean in userformatean],
                                               datanombres=[u'%s %s' % (formatean[0], formatean[1]) for formatean in userformatean],
                                               anchografico=150, altografico=150,
                                               posiciongrafico_x=120, posiciongrafico_y=20,
                                               titulo='SISTEMA OPERATIVO Y APLICATIVOS', tamanotitulo=10,
                                               ubicaciontitulo_x=90, ubicaciontitulo_y=12)
                add_titulo_reportlab(descripcion="_____________________________________________________________________________________________________", tamano=9, alineacion=0, espacios=15, afterespacio=5)
                add_tabla_reportlab(encabezado=[('Nº', 'CAUSAS', 'TOTAL', '%')],
                                    detalles=[(contador1 + 1, liscausas[0], liscausas[1], round(((liscausas[1] * 100)/totalcausas),2)) for contador1, liscausas in enumerate(listadocausas)],
                                    anchocol=[50, 250, 100, 50],
                                    cabecera_left_center=[True, False, True, True],
                                    detalle_left_center=[True, False, True, True])
                add_graficos_circular_reporlab(datavalor=[(lista[1]) for lista in listadocausas],
                                               datanombres=[u'%s %s' % (lista[0], lista[1]) for lista in listadocausas],
                                               anchografico=150, altografico=150,
                                               posiciongrafico_x=120, posiciongrafico_y=20,
                                               titulo='SISTEMA OPERATIVO Y APLICATIVOS', tamanotitulo=10,
                                               ubicaciontitulo_x=90, ubicaciontitulo_y=12)

                add_titulo_reportlab(descripcion="_____________________________________________________________________________________________________", tamano=9, alineacion=0, espacios=15, afterespacio=5)
                add_tabla_reportlab(encabezado=[('Nº', 'ACTIVOS DADOS DE BAJA POR AGENTE', 'TOTAL')],
                                    detalles=[(contador1 + 1, u'%s %s %s'% (lisinfo[0], lisinfo[1], lisinfo[2]), lisinfo[3]) for contador1, lisinfo in enumerate(listadoinformebaja)],
                                    anchocol=[50, 350, 100],
                                    cabecera_left_center=[True, False, True],
                                    detalle_left_center=[True, False, True])
                return generar_pdf_reportlab()
            except Exception as ex:
                return HttpResponseRedirect("/helpdesk_hdincidente?info=%s" % 'Error al generar informe de incidente ')

        elif action == 'seleccionaractivo':
            try:
                if 'id' in request.POST and 'ida' in request.POST:
                    incidente = HdIncidente.objects.get(pk=int(request.POST['id']))
                    activo = ActivoFijo.objects.get(pk=int(request.POST['ida']))
                    incidente.activo = activo
                    incidente.revisionequiposincodigo = False
                    incidente.save(request)
                    log(u'Cambio el código del activo sin coóigo de barra o cogigo interno al incidente: %s' % incidente, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al buacar los datos."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'generarot':
            mensaje = "Problemas al generar Orden de trabajo."
            try:
                listaayudantes = None
                incidente = HdIncidente.objects.get(id=request.POST['id'])
                if incidente.ultimo_registro():
                    listadetalle = incidente.ultimo_registro().hddetalle_incidente_ayudantes_set.values_list('agente_id', flat=True).filter(status=True)
                    listaayudantes = HdDetalle_Grupo.objects.filter(pk__in=listadetalle, status=True)
                if incidente.tipoincidente_id == 3 and incidente.ordentrabajo == None:
                    secuencia = secuencia_ordentrabajo(request, datetime.now().year)
                    secuencia.secuenciaincidente += 1
                    secuencia.save(request)
                    ordentrabajo = OrdenTrabajo(codigoorden="0000"+str(secuencia.secuenciaincidente) + str("-"+str(datetime.now().year)))
                    ordentrabajo.save()
                    incidente.ordentrabajo = ordentrabajo
                    incidente.save()
                # return conviert_html_to_pdf('helpdesk_hdincidente/ordentrabajo.html',
                #                             {'pagesize': 'A4',
                #                              'incidente': incidente,'listaayudantes': listaayudantes,'hoy':datetime.now().date()
                #                              })

                return conviert_html_to_pdf('helpdesk_hdincidente/pdf_incidente.html',
                                            {'pagesize': 'A4',
                                             'incidente': incidente, 'listaayudantes': listaayudantes,
                                             'hoy': datetime.now().date()
                                             })
            except Exception as ex:
                return HttpResponseRedirect("/helpdesk_hdincidente?info=%s" % mensaje)

        elif action == 'imprimirincidente':
            mensaje = "Problemas al Imprimir incidente."
            try:
                listaayudantes = None
                incidente = HdIncidente.objects.get(id=request.POST['id'])
                if incidente.ultimo_registro():
                    listadetalle = incidente.ultimo_registro().hddetalle_incidente_ayudantes_set.values_list('agente_id', flat=True).filter(status=True)
                    listaayudantes = HdDetalle_Grupo.objects.filter(pk__in=listadetalle, status=True)
                materialesincidentes = incidente.hdmaterial_incidente_set.filter(status=True)
                return conviert_html_to_pdf('helpdesk_hdincidente/pdf_incidente.html',
                                            {'pagesize': 'A4',
                                             'incidente': incidente,'listaayudantes': listaayudantes,'materialesincidentes': materialesincidentes,'hoy':datetime.now().date()
                                             })
            except Exception as ex:
                return HttpResponseRedirect("/helpdesk_hdincidente?info=%s" % mensaje)

        elif action=='cerrarorden':
            try:
                form = CerrarOrdenForm(request.POST)
                newfile = None
                if 'archivo' in request.FILES:
                    newfile = request.FILES['archivo']
                    if newfile:
                        if newfile.size > 4194304:
                            return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 4 Mb."})
                        else:
                            newfilesd = newfile._name
                            ext = newfilesd[newfilesd.rfind("."):]
                            if ext == '.pdf' or ext == '.jpg' or ext == '.png' or ext == '.jpeg' or ext == '.PDF' or ext == '.JPG' or ext == '.PNG' or ext == '.JPEG':
                                newfile._name = generar_nombre("ordentrabajo_", newfile._name)
                            else:
                                return JsonResponse({"result": "bad", "mensaje": u"Error, Solo archivo con extención. pdf, jpg, png, jpeg."})
                if form.is_valid():
                    ordentrabajo = OrdenTrabajo.objects.get(pk=int(request.POST['id']))
                    if ordentrabajo.estado != form.cleaned_data['estado']:
                        ordentrabajo.estado = form.cleaned_data['estado']
                    ordentrabajo.informe=form.cleaned_data['informe']
                    ordentrabajo.calificacion=form.cleaned_data['calificacion']
                    ordentrabajo.archivo=newfile
                    ordentrabajo.save(request)
                    datos = json.loads(request.POST['lista_items1'])
                    listaeditar = []
                    if not datos:
                        return JsonResponse({"result": "bad", "mensaje": u"Debe seleccionar personas."})
                    for d in datos:
                        ne = d['ne']
                        if ne[:1] == 'e':
                            listaeditar.append(int(ne[1:ne.__len__()]))
                    ordentrabajo.detalleordentrabajo_set.all().exclude(pk__in=listaeditar).delete()
                    for d in datos:
                        ne = d['ne']
                        repuesto = d['repuesto']
                        cantidad = float(d['cantidad'])
                        if ne[:1] == 'n':
                            detalle = DetalleOrdenTrabajo(orden=ordentrabajo,
                                                          repuesto=repuesto,
                                                          cantidad=cantidad)
                        else:
                            detalle=DetalleOrdenTrabajo.objects.get(pk=int(ne[1:ne.__len__()]))
                            detalle.repuesto=repuesto
                            detalle.cantidad=cantidad
                        detalle.save(request)
                    log(u'Cambio estado de orden: %s' % ordentrabajo, request, "edit")
                    if form.cleaned_data['estado'] == 2:
                        incidente = HdIncidente.objects.filter(status=True, ordentrabajo=ordentrabajo)[0]
                        if incidente.tipoincidente:
                            agente = HdDetalle_Grupo.objects.get(persona=persona, estado=True, grupo__tipoincidente=incidente.tipoincidente)
                        else:
                            agente = HdDetalle_Grupo.objects.get(persona_id=persona.id, estado=True)
                        det = HdDetalle_Incidente(incidente=incidente,
                                                  agente=agente,
                                                  grupo=agente.grupo,
                                                  resolucion=form.cleaned_data['informe'],
                                                  fecharesolucion=datetime.now().date(),
                                                  horaresolucion=datetime.now().time(),
                                                  estadoasignacion=1,
                                                  responsable_id=agente.grupo.hddetalle_grupo_set.get(responsable=True).persona.id,
                                                  estado_id = 3
                                                  )
                        det.save(request)
                        log(u'Resolvio el incidente: %s' % det, request, "edit")
                        if not incidente.tipoincidente:
                            incidente.tipoincidente = agente.grupo.tipoincidente
                        incidente.estado_id = 3
                        incidente.save(request)
                    return JsonResponse({"result": "ok", "mensaje": u"Se cerro correcto."})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'detalle_activo':
            try:
                data['activo'] = activo = ActivoFijo.objects.get(pk=int(request.POST['id']))
                template = get_template("helpdesk_hdincidente/detalle_activo.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'html': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'delorden':
            try:
                orden = OrdenTrabajo.objects.get(pk=int(request.POST['id']))
                log(u'Elimino orden de trabajo: %s' % orden, request, "del")
                orden.status=False
                orden.save()
                return JsonResponse({"result": "ok", "mensaje": u"Se elimino correctamente.."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'aprobarIngresomateriales': #----------------------------------------------------------------------------------------
            try:
                incidente = HdIncidente.objects.get(pk=int(request.POST['id']))
                incidente.aprobarmaterial=True
                incidente.save(request)
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        return JsonResponse({"result": "bad", "mensaje": u"Solicitud Incorrecta."})
    else:
        getNUM = lambda x: int(request.GET.get(x, 0))
        getSTR = lambda x: request.GET.get(x, '')
        if 'action' in request.GET:
            action = request.GET['action']

            if action == 'modaldetalle':
                try:
                    data['incidente'] = incidente = HdIncidente.objects.get(pk=int(encrypt(request.GET['id'])), status=True)
                    if incidente.usuario_creacion:
                        data['personacreacion'] = Persona.objects.get(usuario=incidente.usuario_creacion) if incidente.usuario_creacion.id > 1 else ""

                    if incidente.activo:
                        fecha = datetime.strftime(incidente.activo.fechaingreso, '%Y-%m-%d')
                        vfecha = fecha.split('-')
                        vfecha[0] = str(int(vfecha[0]) + incidente.activo.vidautil)
                        data['fechacaducidad'] = str(vfecha[0] + '/' + vfecha[1] + '/' + vfecha[2])

                    incident = incidente.h_d_incidentes.filter(fecharesolucion__isnull=False, estado__id=3, status=True).first()
                    if incident:
                        resolucion = incident.fecharesolucion
                        data['fechaH'] = incident.horaresolucion.hour - incidente.horareporte.hour
                        data['fechaR'] = (resolucion - incidente.fechareporte).days

                    template = get_template("helpdesk_hdincidente/modaldetalle.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})


            if action == 'modalproductodetalle':
                try:
                    data['incidente'] = incidente = HdIncidente.objects.get(pk=int(encrypt(request.GET['id'])), status=True)
                    # data['proincidente'] = HdIncidenteProductoDetalle.objects.filter(incidente_id = incidente.id)
                    data['productos'] = BodegaKardex.objects.filter(status=True, detalleincidente__incidente_id=int(encrypt(request.GET['id'])))
                    data['hdincidenteprodet'] = HdIncidenteProductoDetalle.objects.filter(status=True, incidente_id=int(encrypt(request.GET['id'])))


                    template = get_template("helpdesk_hdincidente/modalproductodetalle.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            #Funciones
            if action == 'buscaradmin':
                try:
                    q = request.GET['q'].upper().strip()
                    s = q.split(" ")
                    if len(s) == 1:
                        per = Persona.objects.filter((Q(distributivopersona__isnull=False)|Q(profesor__isnull=False)), (Q(nombres__icontains=q)|Q(apellido1__icontains=q)|Q(apellido2__icontains=q)|Q(cedula__contains=q)), Q(status= True)).distinct()[:15]
                    elif len(s) == 2:
                        per = Persona.objects.filter((Q(distributivopersona__isnull=False)|Q(profesor__isnull=False)), (Q(apellido1__contains=s[0]) & Q(apellido2__contains=s[1]))| (Q(nombres__icontains=s[0]) & Q(nombres__icontains=s[1]))|(Q(nombres__icontains=s[0])&Q(apellido1__contains=s[1]))).filter(status=True).distinct()[:15]
                    else:
                        per = Persona.objects.filter((Q(distributivopersona__isnull=False)|Q(profesor__isnull=False)), (Q(nombres__contains=s[0]) & Q(apellido1__contains=s[1]) & Q(apellido2__contains=s[2])) | (Q(nombres__contains=s[0]) & Q(nombres__contains=s[1]) & Q(apellido1__contains=s[2]))).filter(status=True).distinct()[:15]
                    data = {"result": "ok","results": [{"id": x.id, "name": x.flexbox_repr()} for x in per]}
                    # if len(s) == 1:
                    #     admin = Administrativo.objects.filter(Q(persona__nombres__icontains=q)|Q(persona__apellido1__icontains=q)|Q(persona__apellido2__icontains=q)|Q(persona__cedula__contains=q), status= True, activo=True).distinct()[:20]
                    # else:
                    #     admin = Administrativo.objects.filter((Q(persona__apellido1__contains=s[0]) & Q(persona__apellido2__contains=s[1]))).filter(status=True,activo=True ).distinct()[:20]
                    # data = {"result": "ok","results": [{"id": x.persona.id, "name": x.flexbox_repr()} for x in admin]}
                    return JsonResponse(data)
                except Exception as ex:
                    pass

            elif action == 'buscardepartamento':
                try:
                    q = request.GET['q'].upper().strip()
                    s = q.split(" ")
                    if s.__len__() == 2:
                        depar = Departamento.objects.filter(nombre__icontains=s[0],codigo__icontains=s[0],alias__icontains=s[0], status= True).distinct()[:20]
                    else:
                        depar = Departamento.objects.filter((Q(nombre__icontains=s[0]) | Q(codigo__icontains=s[0]) | Q(alias__icontains=s[0]))).filter(status=True).distinct()[:20]
                    data = {"result": "ok","results": [{"id": x.id, "name": x.flexbox_repr()} for x in depar]}
                    return JsonResponse(data)
                except Exception as ex:
                    pass

            elif action == 'buscaractivo':
                try:
                    q = request.GET['q'].upper().strip()
                    s = q.split(" ")
                    tipo = int(request.GET['idt'])
                    if s.__len__() == 1:
                        # activo = ActivoFijo.objects.filter(Q(codigogobierno__icontains=q) | Q(codigointerno__icontains=s[0]) | Q(serie__icontains=q),status=True, catalogo__equipoelectronico=True if tipo == 2 else False).distinct()[:20]
                        activo = ActivoFijo.objects.filter(Q(codigogobierno__icontains=q) | Q(codigointerno__icontains=s[0]) | Q(serie__icontains=q),status=True).distinct()[:20]
                    else:
                        activo = ActivoFijo.objects.filter((Q(codigogobierno__icontains=s[0])& Q(codigointerno__icontains=s[1]))|
                                                           (Q(codigogobierno__icontains=s[0]) & Q(serie__icontains=s[1])) |
                                                           (Q(codigogobierno__icontains=s[0]) & Q(descripcion__icontains=s[1])) |
                                                           (Q(codigointerno__icontains=s[0]) & Q(codigogobierno__icontains=s[1])) |
                                                           (Q(codigointerno__icontains=s[0]) & Q(serie__icontains=s[1])) |
                                                           # (Q(codigointerno__icontains=s[0]) & Q(codigogobierno__icontains=s[1]))).filter(status=True, catalogo__equipoelectronico=True if tipo == 2 else False).distinct()[:20]
                                                           (Q(codigointerno__icontains=s[0]) & Q(codigogobierno__icontains=s[1]))).filter(status=True).distinct()[:20]
                    data = {"result": "ok","results": [{"id": x.id, "name": x.flexbox_reprhd()} for x in activo]}
                    return JsonResponse(data)
                except Exception as ex:
                    pass


            elif action == 'buscaractivotecnologico':
                from sagest.models import ActivoTecnologico
                try:
                    ids = DocumentoFirmaInformeBaja.objects.filter(status=True).values_list('informe__activofijo_id',flat=True)
                    filtro = (Q(status=True,activotecnologico__procesobaja=False,activotecnologico__statusactivo=1,
                                activotecnologico__archivobaja=''))
                    search = request.GET['q'].upper().strip()
                    s = search.split(" ")
                    if s.__len__() == 1:
                        # activo = ActivoFijo.objects.filter(Q(codigogobierno__icontains=q) | Q(codigointerno__icontains=s[0]) | Q(serie__icontains=q),status=True, catalogo__equipoelectronico=True if tipo == 2 else False).distinct()[:20]
                        filtro = filtro & (Q(descripcion__icontains=search) |
                                           Q(activotecnologico__codigogobierno=search) |
                                           Q(serie__icontains=search) |
                                           Q(activotecnologico__codigointerno=search) |
                                           Q(codigotic=search) |
                                           Q(modelo__icontains=search) |
                                           Q(responsable__nombres__icontains=search) |
                                           Q(responsable__apellido1__icontains=search) |
                                           Q(responsable__apellido2__icontains=search))
                    else:
                        filtro = filtro & (Q(responsable__apellido1__unaccent__icontains=s[0]) &
                                           Q(responsable__apellido2__unaccent__icontains=s[1]))

                    activo  = ActivoTecnologico.objects.filter(filtro).exclude(id__in=ids).distinct().order_by('-id')

                    data = {"result": "ok","results": [{"id": x.id, "name": x.flexbox_repr_lic()} for x in activo]}
                    return JsonResponse(data)
                except Exception as ex:
                    pass

            elif action == 'buscaractivogrupo':
                try:
                    q = request.GET['q'].upper().strip()
                    s = q.split(" ")
                    codigogrupo = 0
                    if 'codigoactivo' in request.GET:
                        codigogrupo = request.GET['codigoactivo']
                    if s.__len__() == 1:
                        if codigogrupo == '2':
                            activo = ActivoFijo.objects.filter(Q(codigogobierno__icontains=q) | Q(codigointerno__icontains=s[0]) | Q(serie__icontains=q),status=True,catalogo__equipoelectronico=True).distinct()[:20]
                        else:
                            activo = ActivoFijo.objects.filter(Q(codigogobierno__icontains=q) | Q(codigointerno__icontains=s[0]) | Q(serie__icontains=q),status=True).distinct()[:20]
                    else:
                        if codigogrupo == '2':
                            activo = ActivoFijo.objects.filter(
                                (Q(codigogobierno__icontains=s[0]) & Q(codigointerno__icontains=s[1])) |
                                (Q(codigogobierno__icontains=s[0]) & Q(serie__icontains=s[1])) |
                                (Q(codigogobierno__icontains=s[0]) & Q(descripcion__icontains=s[1])) |
                                (Q(codigointerno__icontains=s[0]) & Q(codigogobierno__icontains=s[1])) |
                                (Q(codigointerno__icontains=s[0]) & Q(serie__icontains=s[1])) |
                                (Q(codigointerno__icontains=s[0]) & Q(codigogobierno__icontains=s[1]))).filter(status=True,catalogo__equipoelectronico=True).distinct()[:20]
                        else:
                            activo = ActivoFijo.objects.filter((Q(codigogobierno__icontains=s[0])& Q(codigointerno__icontains=s[1]))|
                                                               (Q(codigogobierno__icontains=s[0]) & Q(serie__icontains=s[1])) |
                                                               (Q(codigogobierno__icontains=s[0]) & Q(descripcion__icontains=s[1])) |
                                                               (Q(codigointerno__icontains=s[0]) & Q(codigogobierno__icontains=s[1])) |
                                                               (Q(codigointerno__icontains=s[0]) & Q(serie__icontains=s[1])) |
                                                               (Q(codigointerno__icontains=s[0]) & Q(codigogobierno__icontains=s[1]))).filter(status=True).distinct()[:20]
                    data = {"result": "ok","results": [{"id": x.id, "name": x.flexbox_reprhd()} for x in activo]}
                    return JsonResponse(data)
                except Exception as ex:
                    pass

            elif action == 'aprobarIngresomateriales':    #----------------------------------------------------------------------------------------
                try:
                    data['title'] = u'Aprobar el ingreso de Materiales en la Orden'
                    data['data'] = HdIncidente.objects.get(pk=int(request.GET['id']))
                    return render(request, "helpdesk_hdincidente/aprobarmaterialesmodal.html", data)
                except Exception as ex:
                    pass

            elif action == 'addtiempo':
                try:
                    if 'id' in request.GET:
                        prioridad = HdPrioridad.objects.get(pk=int(request.GET['id']))
                        data = {"results": "ok", "hora": prioridad.horamax, "minuto": prioridad.minutomax,"segundo": prioridad.segundomax}
                        return JsonResponse(data)
                except Exception as ex:
                    pass

            elif action == 'busqueda':
                try:
                    q = request.GET['q'].upper().strip()
                    s = q.split(" ")
                    detalle_grupo = HdDetalle_Grupo.objects.values_list('persona_id', flat=True).filter(estado=True,status=True)
                    if s.__len__() == 2:
                        ins = Administrativo.objects.filter(persona__apellido1__icontains=s[0],persona__apellido2__icontains=s[1],status=True).exclude(Q(persona__id=persona.id) | Q(persona_id__in=detalle_grupo, )).distinct()
                    else:
                        ins = Administrativo.objects.filter((Q(persona__nombres__contains=s[0]) | Q(persona__apellido1__contains=s[0]) | Q(persona__apellido2__contains=s[0]) | Q(persona__cedula__contains=s[0]))).filter(status=True).exclude(Q(persona_id=persona.id) | Q(persona_id__in=detalle_grupo, )).distinct()
                    data = {"result": "ok", "results": [{"id": x.persona_id, "name": x.flexbox_repr()} for x in ins]}
                    return JsonResponse(data)
                except Exception as ex:
                    pass
            #Incidente
            elif action == 'add':
                try:
                    data['title'] = u'Adicionar Incidente 123'
                    estado = HdEstado.objects.get(id=2)
                    form = HdIncidenciaFrom(initial={'estado':estado})
                    agente = persona.hddetalle_grupo_set.filter(status=True).first()
                    form.fields['grupo'].queryset = HdGrupo.objects.filter(pk=None)
                    form.fields['agente'].queryset = HdDetalle_Grupo.objects.filter(pk=None)
                    form.fields['subcategoria'].queryset = HdSubCategoria.objects.filter(pk=None)
                    form.fields['bloque'].queryset = HdBloque.objects.filter(status=True, hdbloqueubicacion__bloque__isnull=False).distinct('id')
                    form.fields['ubicacion'].queryset = HdBloqueUbicacion.objects.filter(pk=None)
                    form.fields['detallesubcategoria'].queryset = HdDetalle_SubCategoria.objects.filter(pk=None)
                    form.fields['estado'].queryset = HdEstado.objects.filter(status=True, pk=1)
                    form.fields['tipoincidente'].queryset = HdTipoIncidente.objects.filter(pk=agente.grupo.tipoincidente.id if agente else None)
                    data['tiene_director'] = True if HdDirector.objects.filter(vigente=True).exists() else False
                    form.add()
                    data['form'] = form
                    return render(request, "helpdesk_hdincidente/add.html", data)
                except Exception as ex:
                    pass

            elif action == 'addincidente':
                try:
                    data['title'] = u'Adicionar Incidente'
                    form = IncidenteForm()
                    data['form'] = form
                    return render(request, "helpdesk_hdincidente/addincidente.html", data)
                except Exception as ex:
                    pass

            elif action == 'editincidente':
                try:
                    data['title'] = u'Editar Incidente'
                    incidente = HdIncidente.objects.get(pk=int(encrypt(request.GET['id'])))
                    data['incidente'] = incidente

                    form = IncidenteForm(initial={
                        'tipogrupo': incidente.tipoincidente,
                        'bloque': incidente.ubicacion.bloque,
                        'ubicacion': incidente.ubicacion,
                        'incidente': incidente.asunto,
                        'fecha': incidente.fechareporte,
                        'hora': incidente.horareporte.strftime('%H:%M')
                    })

                    form.editar(incidente)
                    data['form'] = form
                    return render(request, "helpdesk_hdincidente/editincidente.html", data)
                except Exception as ex:
                    pass

            elif action == 'anularresolucion':
                try:
                    data['title'] = u'Anular Resolución de Incidente'
                    data['incidente'] = HdIncidente.objects.get(pk=int(encrypt(request.GET['id'])))

                    template = get_template("helpdesk_hdincidente/anularresolucion.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content, 'title': data['title']})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            elif action == 'addsalidalicenciasoft':
                try:
                    data['title'] = u'Adicionar entrega de Licencia'
                    form = HdIncidenciaLicenciaFrom(initial={'asunto': 'Instalación de Licencia de software',
                                                             'resolucion': 'Licencia de software Instalada'})

                    form2 = HdDetalleIncidenciaLicenciaFrom()
                    # pro_stock = BodegaKardex.filter(status=True).order_by('producto__id').values('producto__id').distinct()
                    pro_stock = BodegaProducto.objects.filter(status=True, grupo_id=26)
                    listpro = []
                    for pro in pro_stock:
                        if pro.bodegakardex_set.filter(status=True):
                            if pro.cantidad_total() > 0:
                                listpro.append(pro.id)
                    form2.fields['producto'].queryset = BodegaProducto.objects.filter(id__in=listpro)
                    data['form'] = form
                    data['form2'] = form2
                    return render(request, "helpdesk_hdincidente/addlicenciasoftware.html", data)
                except Exception as ex:
                    pass

            elif action == 'edit':
                try:
                    data['title'] = u'Editar Incidente'
                    data['incidente'] = incidente = HdIncidente.objects.get(pk=int(request.GET['id']))
                    grupo = None
                    agente = None
                    ayudantes = None
                    if incidente.h_d_incidentes.filter(status=True).exists():
                        ayudantes = HdDetalle_Grupo.objects.filter(pk__in=HdDetalle_Incidente_Ayudantes.objects.values_list('agente__id').filter(detallleincidente=incidente.ultimo_registro().id, status=True), status=True)
                        if incidente.ultimo_registro().grupo:
                            grupo = incidente.ultimo_registro().grupo
                        # if incidente.ultimo_registro().gr:
                        #     grupo = incidente.ultimo_registro().grupo
                        if incidente.ultimo_registro().agente:
                            agente = incidente.ultimo_registro().agente
                    # else:
                    #     agente = HdDetalle_Grupo.objects.filter(persona=persona, grupo__tipoincidente=incidente.tipoincidente)[0]
                    #     grupo = agente.grupo
                    if incidente.detallesubcategoria:
                        categoria=incidente.detallesubcategoria.subcategoria.categoria
                        subcategoria = incidente.detallesubcategoria.subcategoria
                        data['categoria'] = categoria.id
                    elif incidente.subcategoria :
                        categoria =incidente.subcategoria.categoria
                        subcategoria = incidente.subcategoria
                        data['categoria'] = categoria.id
                    else:
                        categoria=None
                        subcategoria=None
                        data['categoria'] = 0
                    form = HdIncidenciaFrom(initial={'asunto': incidente.asunto,
                                                     'persona': incidente.persona.id,
                                                     'grupo': grupo,
                                                     'agente': agente,
                                                     'categoria': categoria,
                                                     'subcategoria': subcategoria,
                                                     # 'detallesubcategoria': incidente.detallesubcategoria if incidente.detallesubcategoria else None,
                                                     'activo': incidente.activo.id if incidente.activo else 0,
                                                     'fechacompra': incidente.activo.fechaingreso if incidente.activo else None,
                                                     'vidautil': incidente.activo.vidautil if incidente.activo else None,
                                                     'tiemporestante': incidente.activo.fecha_caducidad() if incidente.activo else None,
                                                     'medioreporte': incidente.medioreporte,
                                                     'fechareporte': incidente.fechareporte.strftime('%d-%m-%Y'),
                                                     'horareporte':incidente.horareporte.strftime('%H-%M-%S'),
                                                     'estado': incidente.estado,
                                                     'bloque': incidente.ubicacion.bloque if incidente.ubicacion else None,
                                                     'ubicacion': incidente.ubicacion if incidente.ubicacion else None,
                                                     'tipoincidente': incidente.tipoincidente if incidente.tipoincidente else agente.grupo.tipoincidente,
                                                     'revisionequipoexterno':incidente.revisionequipoexterno,
                                                     'revisionequiposincodigo':incidente.revisionequiposincodigo,
                                                     'serie':incidente.serie,
                                                     'ayudantes':ayudantes
                                                     })
                    if grupo:
                        form.fields['agente'].queryset = grupo.mis_agentes()
                    form.fields['grupo'].queryset = HdTipoIncidente.objects.filter(status=True,pk=grupo.tipoincidente_id) if grupo else HdGrupo.objects.filter(status=True,tipoincidente_id=int(incidente.tipoincidente_id))
                    # if incidente.detallesubcategoria:
                    if incidente.detallesubcategoria:
                        form.fields['categoria'].queryset = HdCategoria.objects.filter(tipoincidente_id=incidente.detallesubcategoria.subcategoria.categoria.tipoincidente_id)
                    # else:
                    #     form.fields['categoria'].queryset = HdCategoria.objects.filter(pk=None)
                    form.fields['ubicacion'].queryset = HdBloqueUbicacion.objects.filter(status=True, bloque=incidente.ubicacion.bloque).order_by('id') if incidente.ubicacion else None
                    form.fields['persona'].widget.attrs['descripcion'] = incidente.persona
                    form.fields['persona'].widget.attrs['value'] = incidente.persona.id
                    if grupo:
                        form.fields['agente'].queryset = HdDetalle_Grupo.objects.filter(grupo_id=incidente.ultimo_registro().grupo.id, estado=True)
                    else:
                        form.fields['agente'].queryset =  HdDetalle_Grupo.objects.filter(pk=None)
                    if incidente.activo:
                        form.fields['activo'].widget.attrs['descripcion'] = incidente.activo
                        form.fields['activo'].widget.attrs['value'] = incidente.activo.id
                    else:
                        form.fields['activo'].widget.attrs['descripcion'] = '----------------'
                        form.fields['activo'].widget.attrs['value'] = 0
                    if incidente.detallesubcategoria:
                        form.fields['subcategoria'].queryset = HdSubCategoria.objects.filter(pk=incidente.detallesubcategoria.subcategoria_id)
                        form.fields['detallesubcategoria'].queryset = HdDetalle_SubCategoria.objects.filter(pk=incidente.detallesubcategoria_id)
                    else:
                        form.fields['subcategoria'].queryset = HdSubCategoria.objects.filter(pk=incidente.subcategoria_id)
                        form.fields['detallesubcategoria'].queryset = HdDetalle_SubCategoria.objects.filter(pk=None)
                    form.editar(incidente)
                    if incidente.mi_detalle():
                        if incidente.ultimo_registro().agente:
                            form.tiene_agente()
                    if incidente.tipoincidente.id == 3:
                        data['nova'] = 1
                    else:
                        data['nova'] = 2
                    data['form'] = form
                    return render(request, "helpdesk_hdincidente/edit.html", data)
                except Exception as ex:
                    pass

            elif action == 'editdetalle':
                try:
                    data['title'] = u'Editar detalle de Incidente'
                    data['detalle'] = detalle = HdDetalle_Incidente.objects.get(pk=int(request.GET['id']))
                    data['form'] = HdDetalleIncidenteFrom(initial={'grupo': detalle.grupo,
                                                                   'agente': detalle.agente,
                                                                   'estado': detalle.estado
                                                                   })
                    return render(request, "helpdesk_hdincidente/editdetalle.html", data)
                except Exception as ex:
                    pass

            # if action == 'modaldetalle':
            #     try:
            #         if 'id' in request.GET:
            #             data['incidente'] =incidente= HdIncidente.objects.get(pk=int(request.GET['id']), status=True)
            #             data['puede_vizualisar_detalle'] = True if incidente.mi_detalle().count() > 1 else False
            #             if incidente.usuario_creacion:
            #                 data['personacreacion'] = Persona.objects.get(usuario=incidente.usuario_creacion) if incidente.usuario_creacion.id > 1 else ""
            #             if incidente.activo:
            #                 fecha = datetime.strftime(incidente.activo.fechaingreso, '%Y-%m-%d')
            #                 vfecha = fecha.split('-')
            #                 vfecha[0] = str(int(vfecha[0]) + incidente.activo.vidautil)
            #                 data['fechacaducidad'] = str(vfecha[0] + '/' + vfecha[1] + '/' + vfecha[2])
            #             template = get_template("helpdesk_hdincidente/modaldetalle.html")
            #             json_content = template.render(data)
            #             return JsonResponse({"result": "ok", 'data': json_content})
            #     except Exception as ex:
            #         transaction.set_rollback(True)
            #         return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            elif action == 'reasignaragente':
                try:
                    data['title'] = u'Reasignar Agente'
                    data['incidente'] = incidente = HdIncidente.objects.get(pk=int(request.GET['id']))
                    grupo = None
                    agente = None
                    if incidente.h_d_incidentes.filter(status=True).exists():
                        if incidente.ultimo_registro().estadoasignacion == 1:
                            if incidente.ultimo_registro().grupo:
                                grupo = incidente.ultimo_registro().grupo
                            if incidente.ultimo_registro().agente:
                                agente = incidente.ultimo_registro().agente
                        else:
                            agente = HdDetalle_Grupo.objects.filter(persona=persona, grupo__tipoincidente=incidente.tipoincidente)[0]
                            grupo = agente.grupo
                    else:
                        agente = HdDetalle_Grupo.objects.filter(persona=persona, grupo__tipoincidente=incidente.tipoincidente)[0]
                        grupo = agente.grupo
                    form = HdIncidenciaFrom(initial={'asunto': incidente.asunto,
                                                     'fechareporte': incidente.fechareporte,
                                                     'persona': incidente.persona.id,
                                                     'categoria': incidente.detallesubcategoria.subcategoria.categoria if incidente.detallesubcategoria else None,
                                                     'subcategoria': incidente.detallesubcategoria.subcategoria if incidente.detallesubcategoria else None,
                                                     'detallesubcategoria': incidente.detallesubcategoria if incidente.detallesubcategoria else None,
                                                     'prioridad': incidente.detallesubcategoria.prioridad.prioridad.nombre if incidente.detallesubcategoria else None,
                                                     'activo': incidente.activo.id if incidente.activo else 0,
                                                     'fechacompra': incidente.activo.fechaingreso if incidente.activo else None,
                                                     'vidautil': incidente.activo.vidautil if incidente.activo else None,
                                                     'tiemporestante': incidente.activo.fecha_caducidad() if incidente.activo else None,
                                                     'descripcion': incidente.descripcion,
                                                     'estado': HdEstado.objects.get(pk=2),
                                                     'grupo': grupo,
                                                     'bloque': incidente.ubicacion.bloque if incidente.ubicacion else None,
                                                     'ubicacion': incidente.ubicacion if incidente.ubicacion else None,
                                                     'tipoincidente': incidente.tipoincidente if incidente.tipoincidente else agente.grupo.tipoincidente,
                                                     'medioreporte': incidente.medioreporte,
                                                     'horareporte': incidente.horareporte,
                                                     'revisionequipoexterno': incidente.revisionequipoexterno,
                                                     'revisionequiposincodigo': incidente.revisionequiposincodigo,
                                                     'serie': incidente.serie
                                                     })
                    es_tics = False
                    if grupo:
                        form.fields['agente'].queryset = grupo.mis_agentes()
                        if grupo.tipoincidente.id == 2:
                            es_tics = True
                    data['es_tics'] = es_tics
                    form.fields['grupo'].queryset = grupo.tipoincidente.t_incidente.filter(status=True)
                    form.fields['persona'].widget.attrs['descripcion'] = incidente.persona
                    form.fields['persona'].widget.attrs['value'] = incidente.persona.id
                    form.fields['categoria'].queryset = HdCategoria.objects.filter(tipoincidente=grupo.tipoincidente, status=True)
                    if incidente.activo:
                        form.fields['activo'].widget.attrs['descripcion'] = incidente.activo
                        form.fields['activo'].widget.attrs['value'] = incidente.activo.id
                    else:
                        form.fields['activo'].widget.attrs['descripcion'] = '----------------'
                        form.fields['activo'].widget.attrs['value'] = 0
                    if incidente.detallesubcategoria:
                        form.fields['subcategoria'].queryset = HdSubCategoria.objects.filter(pk=incidente.detallesubcategoria.subcategoria_id)
                        form.fields['detallesubcategoria'].queryset = HdDetalle_SubCategoria.objects.filter(pk=incidente.detallesubcategoria_id)
                    else:
                        form.fields['subcategoria'].queryset = HdSubCategoria.objects.filter(pk=None)
                        form.fields['detallesubcategoria'].queryset = HdDetalle_SubCategoria.objects.filter(pk=None)
                    form.reasignar()
                    form.cargarayudantes(grupo)
                    data['form'] = form
                    return render(request, "helpdesk_hdincidente/reasignargente.html", data)
                except Exception as ex:
                    pass

            elif action == 'escalarincidente':
                try:
                    data['title'] = u'Escalar Incidente a otro Grupo'
                    data['incidente'] = incidente = HdIncidente.objects.get(pk=int(request.GET['id']))
                    agente = None
                    if incidente.h_d_incidentes.filter(status=True).exists():
                        if incidente.ultimo_registro().agente:
                            agente = incidente.ultimo_registro().agente
                    else:
                        agente = HdDetalle_Grupo.objects.filter(persona=persona, grupo__tipoincidente=incidente.tipoincidente)[0]
                    form = HdIncidenciaFrom(initial={'asunto': incidente.asunto,
                                                     'fechareporte': incidente.fechareporte,
                                                     'persona': incidente.persona.id,
                                                     'categoria': incidente.detallesubcategoria.subcategoria.categoria if incidente.detallesubcategoria else None,
                                                     'subcategoria': incidente.detallesubcategoria.subcategoria if incidente.detallesubcategoria else None,
                                                     'detallesubcategoria': incidente.detallesubcategoria if incidente.detallesubcategoria else None,
                                                     'prioridad': incidente.detallesubcategoria.prioridad.prioridad.nombre if incidente.detallesubcategoria else None,
                                                     'activo': incidente.activo.id if incidente.activo else 0,
                                                     'fechacompra': incidente.activo.fechaingreso if incidente.activo else None,
                                                     'vidautil': incidente.activo.vidautil if incidente.activo else None,
                                                     'tiemporestante': incidente.activo.fecha_caducidad() if incidente.activo else None,
                                                     'descripcion': incidente.descripcion,
                                                     'bloque': incidente.ubicacion.bloque if incidente.ubicacion else None,
                                                     'ubicacion': incidente.ubicacion if incidente.ubicacion else None,
                                                     'tipoincidente': incidente.tipoincidente if incidente.tipoincidente else agente.grupo.tipoincidente,
                                                     'medioreporte': incidente.medioreporte,
                                                     'horareporte': incidente.horareporte
                                                     })
                    form.fields['persona'].widget.attrs['descripcion'] = incidente.persona
                    form.fields['persona'].widget.attrs['value'] = incidente.persona.id
                    es_tics = False
                    if incidente.tipoincidente.id == 2:
                        es_tics = True
                    data['es_tics'] = es_tics
                    if incidente.activo:
                        form.fields['activo'].widget.attrs['descripcion'] = incidente.activo
                        form.fields['activo'].widget.attrs['value'] = incidente.activo.id
                    else:
                        form.fields['activo'].widget.attrs['descripcion'] = '----------------'
                        form.fields['activo'].widget.attrs['value'] = 0
                    if incidente.detallesubcategoria:
                        form.fields['subcategoria'].queryset = HdSubCategoria.objects.filter(pk=incidente.detallesubcategoria.subcategoria_id)
                        form.fields['detallesubcategoria'].queryset = HdDetalle_SubCategoria.objects.filter(pk=incidente.detallesubcategoria_id)
                    else:
                        form.fields['subcategoria'].queryset = HdSubCategoria.objects.filter(pk=None)
                        form.fields['detallesubcategoria'].queryset = HdDetalle_SubCategoria.objects.filter(pk=None)
                    form.escalar()
                    data['form'] = form
                    return render(request, "helpdesk_hdincidente/escalarincidente.html", data)
                except Exception as ex:
                    pass

            elif action == 'resolverincidente':
                try:
                    data['title'] = u'Resolver Incidente'
                    data['incidente'] = incidente = HdIncidente.objects.get(pk=int(request.GET['id']))
                    grupo = None
                    agente = None
                    ayudantes = None
                    if incidente.h_d_incidentes.filter(status=True).exists():
                        ayudantes = HdDetalle_Grupo.objects.filter(pk__in=HdDetalle_Incidente_Ayudantes.objects.values_list('agente__id').filter(detallleincidente=incidente.ultimo_registro().id, status=True), status=True)
                        if incidente.ultimo_registro().estadoasignacion == 1:
                            if incidente.ultimo_registro().grupo:
                                grupo = incidente.ultimo_registro().grupo
                            if incidente.ultimo_registro().agente:
                                agente = incidente.ultimo_registro().agente
                        else:
                            agente = HdDetalle_Grupo.objects.filter(persona=persona, grupo__tipoincidente=incidente.tipoincidente)[0]
                            grupo = agente.grupo
                    else:
                        agente = HdDetalle_Grupo.objects.filter(persona=persona, grupo__tipoincidente=incidente.tipoincidente)[0]
                        grupo = agente.grupo
                    form = HdIncidenciaFrom(initial={'asunto': incidente.asunto,
                                                     'fechareporte': incidente.fechareporte,
                                                     'persona': incidente.persona.id,
                                                     'categoria': incidente.detallesubcategoria.subcategoria.categoria if incidente.detallesubcategoria else None,
                                                     'subcategoria': incidente.detallesubcategoria.subcategoria if incidente.detallesubcategoria else None,
                                                     'detallesubcategoria': incidente.detallesubcategoria if incidente.detallesubcategoria else None,
                                                     'prioridad': incidente.detallesubcategoria.prioridad.prioridad.nombre if incidente.detallesubcategoria else None,
                                                     'activo': incidente.activo.id if incidente.activo else 0,
                                                     'fechacompra': incidente.activo.fechaingreso if incidente.activo else None,
                                                     'vidautil': incidente.activo.vidautil if incidente.activo else None,
                                                     'tiemporestante': incidente.activo.fecha_caducidad() if incidente.activo else None,
                                                     'descripcion': incidente.descripcion,
                                                     'estado': incidente.estado,
                                                     'agente': agente,
                                                     'grupo': grupo,
                                                     'bloque': incidente.ubicacion.bloque if incidente.ubicacion else None,
                                                     'ubicacion': incidente.ubicacion if incidente.ubicacion else None,
                                                     'tipoincidente': incidente.tipoincidente if incidente.tipoincidente else agente.grupo.tipoincidente,
                                                     'medioreporte': incidente.medioreporte,
                                                     'horareporte': incidente.horareporte,
                                                     'causa': incidente.causa,
                                                     'revisionequipoexterno': incidente.revisionequipoexterno,
                                                     'revisionequiposincodigo': incidente.revisionequiposincodigo,
                                                     'serie': incidente.serie,
                                                     'ayudantes': ayudantes
                                                     })
                    form.resolver()
                    if HdCausas.objects.filter(tipoincidente=incidente.tipoincidente,status=True).exists():
                        form.fields['causa'].queryset = HdCausas.objects.filter(tipoincidente=incidente.tipoincidente,status=True)
                    form.fields['estado'].queryset = HdEstado.objects.all().exclude(Q(id=1) | Q(id=4))
                    form.fields['persona'].widget.attrs['descripcion'] = incidente.persona
                    form.fields['persona'].widget.attrs['value'] = incidente.persona.id
                    form.fields['categoria'].queryset = HdCategoria.objects.filter(tipoincidente=grupo.tipoincidente, status=True)
                    # form.fields['departamento'].widget.attrs['descripcion'] = incidente.departamento if incidente.departamento else '------------------'
                    # form.fields['departamento'].widget.attrs['value'] = incidente.departamento.id if incidente.departamento else 0
                    es_tics = False
                    if grupo:
                        if grupo.tipoincidente.id == 2:
                            es_tics = True
                    data['es_tics'] = es_tics
                    if incidente.activo:
                        form.fields['activo'].widget.attrs['descripcion'] = incidente.activo
                        form.fields['activo'].widget.attrs['value'] = incidente.activo.id
                    else:
                        form.fields['activo'].widget.attrs['descripcion'] = '----------------'
                        form.fields['activo'].widget.attrs['value'] = 0
                    if incidente.detallesubcategoria:
                        form.fields['subcategoria'].queryset = HdSubCategoria.objects.filter(pk=incidente.detallesubcategoria.subcategoria_id)
                        form.fields['detallesubcategoria'].queryset = HdDetalle_SubCategoria.objects.filter(pk=incidente.detallesubcategoria_id)
                    else:
                        form.fields['subcategoria'].queryset = HdSubCategoria.objects.filter(pk=None)
                        form.fields['detallesubcategoria'].queryset = HdDetalle_SubCategoria.objects.filter(pk=None)

                    data['lista'] = incidente.mi_detalle()

                    data['ayudantes'] = ayudantes

                    resuelto = False
                    if incidente.esta_resulto():
                        resuelto = True
                    data['resuelto'] = resuelto
                    if incidente.tipoincidente.id == 3:
                        data['materialesincidentes'] = incidente.hdmaterial_incidente_set.filter(status=True)
                        data['formmaterialesincidentes'] = HdMaterial_IncidenteForm()
                    data['fechanow'] = datetime.now()
                    data['form'] = form
                    #formulariobodegaproductos
                    form2 = HdDetalleIncidenciaProductoFrom()
                    # pro_stock = BodegaKardex.filter(status=True).order_by('producto__id').values('producto__id').distinct()
                    pro_stock = BodegaProducto.objects.filter(status=True).exclude(grupo_id=26)
                    listpro = []
                    for pro in pro_stock:
                        if pro.bodegakardex_set.filter(status=True):
                            if pro.cantidad_total() > 0:
                                listpro.append(pro.id)
                    form2.fields['producto'].queryset = BodegaProducto.objects.filter(id__in=listpro)
                    data['form2'] = form2
                    return render(request, "helpdesk_hdincidente/resolverincidente.html", data)
                except Exception as ex:
                    pass

            elif action == 'delincidente':
                try:
                    data['title'] = u'Eliminar incidente'
                    data['incidente'] = HdIncidente.objects.get(pk=int(request.GET['id']))
                    return render(request, "helpdesk_hdincidente/delincidente.html", data)
                except Exception as ex:
                    pass
            #Estado
            elif action == 'addestado':
                try:
                    data['title'] = u'Adicionar estado de Help Desk'
                    data['form'] = HdEstadoForm()
                    return render(request, "helpdesk_hdincidente/addestado.html", data)
                except Exception as ex:
                    pass

            elif action=='delestado':
                try:
                    data['title'] = u'Eliminar Estado de Help Desk'
                    data['estado'] = HdEstado.objects.get(pk=int(request.GET['id']))
                    return render(request, "helpdesk_hdincidente/delestado.html", data)
                except Exception as ex:
                    pass

            elif action == 'editestado':
                try:
                    data['title'] = u'Editar Urgencia'
                    data['estado'] = estado = HdEstado.objects.get(pk=int(request.GET['id']))
                    form = HdEstadoEditForm(initial={'nombre':estado.nombre})
                    # if estado.esta_activo():
                    #     form.editar()
                    data['form'] = form
                    return render(request, "helpdesk_hdincidente/editestado.html", data)
                except Exception as ex:
                    pass

            elif action == 'cambiarimagen':
                try:
                    data['title'] = u'Cambiar imagen de estado Help Desk'
                    data['estado'] = HdEstado.objects.get(pk=int(request.GET['id']))
                    form = HdEstadoImagenForm()
                    data['form'] = form
                    return render(request, "helpdesk_hdincidente/cambiarimagen.html", data)
                except Exception as ex:
                    pass

            elif action == 'estado':
                try:
                    data['title'] = u'Estados de Help Desk'
                    # if request.user.has_perm("sagest.puede_ingresar_configurar_helpdesk"):
                    data['estados'] = HdEstado.objects.filter(status=True).order_by('id')
                    return render(request, "helpdesk_hdincidente/estado.html", data)
                except Exception as ex:
                    pass
            # Impacto
            elif action == 'addimpacto':
                try:
                    data['title'] = u'Adicionar Impacto'
                    data['form'] = HdImpactoForm()
                    return render(request, "helpdesk_hdincidente/addimpacto.html", data)
                except Exception as ex:
                    pass

            elif action =='delimpacto':
                try:
                    data['title'] = u'Eliminar Impacto'
                    data['impacto'] = HdImpacto.objects.get(pk=int(request.GET['id']))
                    return render(request, "helpdesk_hdincidente/delimpacto.html", data)
                except Exception as ex:
                    pass

            elif action == 'editimpacto':
                try:
                    data['title'] = u'Editar Impacto'
                    data['impacto'] = impacto=HdImpacto.objects.get(pk=int(request.GET['id']))
                    form = HdUrgenciaForm(initial={'nombre':impacto.nombre,'codigo':impacto.codigo, 'descripcion':impacto.descripcion})
                    if impacto.esta_activo():
                        form.editar()
                    data['form'] = form
                    return render(request, "helpdesk_hdincidente/editimpacto.html", data)
                except Exception as ex:
                    pass

            elif action == 'impacto':
                try:
                    data['title'] = u'Lista de Impacto'
                    # if request.user.has_perm("sagest.puede_ingresar_configurar_helpdesk"):
                    data['impacto'] = HdImpacto.objects.filter(status=True)
                    return render(request, "helpdesk_hdincidente/impacto.html", data)
                except Exception as ex:
                    pass
            #Urgencia
            elif action == 'addurgencia':
                try:
                    data['title'] = u'Adicionar Urgencia'
                    data['form'] = HdUrgenciaForm()
                    return render(request, "helpdesk_hdincidente/addurgencia.html", data)
                except Exception as ex:
                    pass

            elif action == 'addsolicitudpiezaparte':
                try:
                    data['title'] = u'Adicionar Solicitud de Piezas y Partes'
                    data['form'] = HdSolicitudPiezaPartesForm()
                    return render(request, "helpdesk_hdincidente/addsolicitudpiezaparte.html", data)
                except Exception as ex:
                    pass

            elif action == 'addmaterial':
                try:
                    data['title'] = u'Adicionar Material'
                    data['form'] = HdMaterialForm()
                    return render(request, "helpdesk_hdincidente/addmaterial.html", data)
                except Exception as ex:
                    pass

            elif action == 'editmaterial':
                try:
                    data['title'] = u'Editar Material'
                    data['material'] = material = HdMateriales.objects.get(pk=int(request.GET['id']))
                    form = HdMaterialForm(initial={'codigo': material.codigo,
                                                   'nombre': material.nombre })
                    data['form'] = form
                    return render(request, "helpdesk_hdincidente/editmaterial.html", data)
                except Exception as ex:
                    pass

            elif action=='delmaterial':
                try:
                    data['title'] = u'Eliminar Material'
                    data['material'] = HdMateriales.objects.get(pk=int(request.GET['id']))
                    return render(request, "helpdesk_hdincidente/delmaterial.html", data)
                except Exception as ex:
                    pass

            elif action == 'addcausa':
                try:
                    data['title'] = u'Adicionar Causa'
                    data['form'] = HdCausasForm()
                    return render(request, 'helpdesk_hdincidente/addcausa.html', data)
                except Exception as ex:
                    pass

            elif action == 'addfechacierre':
                try:
                    data['title'] = u'Adicionar Fecha cierre'
                    data['form'] = HdFechacierreForm()
                    return render(request, 'helpdesk_hdincidente/addfechacierre.html', data)
                except Exception as ex:
                    pass

            elif action == 'editfechacierre':
                try:
                    data['title'] = u'Editar Fecha cierre'
                    data['fechacierre'] = fechacierre = HdFechacierresolicitudes.objects.get(pk=int(request.GET['id']))
                    form = HdFechacierreForm(initial={ 'observacion':fechacierre.observacion,
                                                       'fechainicio':fechacierre.fechainicio,
                                                       'activo':fechacierre.activo,
                                                       'fechafin': fechacierre.fechafin})
                    data['form'] = form
                    return render(request, "helpdesk_hdincidente/editfechacierre.html", data)
                except Exception as ex:
                    pass

            elif action=='cerrarfechacierre':
                try:
                    data['title'] = u'Cerrar Fecha cierre'
                    data['fechacierre'] = HdFechacierresolicitudes.objects.get(pk=int(request.GET['id']))
                    return render(request, 'helpdesk_hdincidente/cerrarfechacierre.html', data)
                except Exception as ex:
                    pass

            elif action=='delfechacierre':
                try:
                    data['title'] = u'Eliminar Fecha cierre'
                    data['fechacierre'] = HdFechacierresolicitudes.objects.get(pk=int(request.GET['id']))
                    return render(request, 'helpdesk_hdincidente/delfechacierre.html', data)
                except Exception as ex:
                    pass

            elif action == 'addprecio':
                try:
                    data['title'] = u'Adicionar Precio'
                    data['idsolicitud'] = int(request.GET['idsolicitud'])
                    data['form'] = HdPreciosForm()
                    return render(request, 'helpdesk_hdincidente/addprecio.html', data)
                except Exception as ex:
                    pass

            elif action == 'editprecio':
                try:
                    data['title'] = u'Editar Precio'
                    data['precio'] = precio = HdPrecioSolicitudesPiezaPartes.objects.get(pk=int(request.GET['id']))
                    form = HdPreciosForm(initial={ 'valor':precio.valor,'activo':precio.activo,'cierresolicitudes':precio.cierresolicitudes })
                    data['form'] = form
                    return render(request, "helpdesk_hdincidente/ediprecio.html", data)
                except Exception as ex:
                    pass

            elif action=='delprecio':
                try:
                    data['title'] = u'Eliminar Precio'
                    data['precio'] = HdPrecioSolicitudesPiezaPartes.objects.get(pk=int(request.GET['id']))
                    return render(request, "helpdesk_hdincidente/delprecio.html", data)
                except Exception as ex:
                    pass

            elif action == 'editcausa':
                try:
                    data['title'] = u'Editar Causa'
                    data['causa'] = causa = HdCausas.objects.get(pk=int(request.GET['id']))
                    form = HdCausasForm(initial={ 'nombre':causa.nombre,'tipoincidente':causa.tipoincidente })
                    data['form'] = form
                    return render(request, "helpdesk_hdincidente/editcausa.html", data)
                except Exception as ex:
                    pass

            elif action=='delcausa':
                try:
                    data['title'] = u'Eliminar Causa'
                    data['causa'] = HdCausas.objects.get(pk=int(request.GET['id']))
                    return render(request, "helpdesk_hdincidente/delcausa.html", data)
                except Exception as ex:
                    pass

            elif action == 'addpiezaparte':
                try:
                    data['title'] = u'Adicionar Piezas y Partes'
                    data['form'] = HdPiezaPartesForm()
                    return render(request, "helpdesk_hdincidente/addpiezaparte.html", data)
                except Exception as ex:
                    pass

            elif action=='delpiezaparte':
                try:
                    data['title'] = u'Eliminar Pieza y Parte'
                    data['piezaparte'] = HdPiezaPartes.objects.get(pk=int(request.GET['id']))
                    return render(request, "helpdesk_hdincidente/delpiezaparte.html", data)
                except Exception as ex:
                    pass

            elif action=='delsolicitudpiezaparte':
                try:
                    data['title'] = u'Eliminar Solicitud Pieza y Parte'
                    data['solicitudpiezaparte'] = HdSolicitudesPiezaPartes.objects.get(pk=int(request.GET['id']))
                    return render(request, "helpdesk_hdincidente/delsolicitudpiezaparte.html", data)
                except Exception as ex:
                    pass

            if action=='delurgencia':
                try:
                    data['title'] = u'Eliminar Urgencia'
                    data['urgencia'] = HdUrgencia.objects.get(pk=int(request.GET['id']))
                    return render(request, "helpdesk_hdincidente/delurgencia.html", data)
                except Exception as ex:
                    pass

            elif action == 'editurgencia':
                try:
                    data['title'] = u'Editar Urgencia'
                    data['urgencia'] = urgencia=HdUrgencia.objects.get(pk=int(request.GET['id']))
                    form = HdUrgenciaForm(initial={'nombre':urgencia.nombre,'codigo':urgencia.codigo,'descripcion':urgencia.descripcion})
                    if urgencia.esta_activo():
                        form.editar()
                    data['form'] = form
                    return render(request, "helpdesk_hdincidente/editurgencia.html", data)
                except Exception as ex:
                    pass

            elif action == 'urgencia':
                try:
                    data['title'] = u'Lista de Ugencias'
                    # if request.user.has_perm("sagest.puede_ingresar_configurar_helpdesk"):
                    data['ugencia'] = HdUrgencia.objects.filter(status=True)
                    return render(request, "helpdesk_hdincidente/urgencia.html", data)
                except Exception as ex:
                    pass

            elif action == 'listadocausas':
                try:
                    data['title'] = u'Causas de Incidentes'
                    data['listadocausas'] = HdCausas.objects.filter(status=True)
                    return render(request, "helpdesk_hdincidente/listadocausas.html", data)
                except Exception as ex:
                    pass

            elif action == 'listadoprecios':
                try:
                    data['title'] = u'Precios de pieza y parte'
                    data['idsolicitud'] = int(request.GET['id'])
                    data['listadoprecios'] = HdPrecioSolicitudesPiezaPartes.objects.filter(solicitudes_id=int(request.GET['id']),status=True).order_by('-id')
                    return render(request, "helpdesk_hdincidente/listadoprecios.html", data)
                except Exception as ex:
                    pass

            elif action == 'listadomantenimientofecha':
                try:
                    data['title'] = u'Fechas cortes'
                    data['listadomantenimientofecha'] = HdFechacierresolicitudes.objects.filter(status=True).order_by('-id')
                    return render(request, "helpdesk_hdincidente/listadomantenimientofecha.html", data)
                except Exception as ex:
                    pass

            elif action == 'piezapartes':
                try:
                    data['title'] = u'Piezas y Partes'
                    data['piezapartes'] = HdPiezaPartes.objects.filter(status=True)
                    return render(request, "helpdesk_hdincidente/piezapartes.html", data)
                except Exception as ex:
                    pass

            elif action == 'editpiezapartes':
                try:
                    data['title'] = u'Editar Piezas y Partes'
                    data['piezapartes'] = piezapartes=HdPiezaPartes.objects.get(pk=int(request.GET['id']))
                    form = HdPiezaPartesForm(initial={'descripcion':piezapartes.descripcion,'estado':piezapartes.estado})
                    data['form'] = form
                    return render(request, "helpdesk_hdincidente/editpiezapartes.html", data)
                except Exception as ex:
                    pass

            elif action == 'editsolicitudespiezapartes':
                try:
                    data['title'] = u'Editar Solicitudes Piezas y Partes'
                    data['solicitudpiezapartes'] = solicitudpiezapartes = HdSolicitudesPiezaPartes.objects.get(pk=int(request.GET['id']))
                    form = HdSolicitudPiezaPartesForm(initial={'piezaparte':solicitudpiezapartes.piezaparte,
                                                               'grupocategoria':solicitudpiezapartes.grupocategoria,
                                                               'capacidad':solicitudpiezapartes.capacidad,
                                                               'velocidad':solicitudpiezapartes.velocidad,
                                                               'descripcion':solicitudpiezapartes.descripcion,
                                                               'tipo':solicitudpiezapartes.tipo })
                    data['form'] = form
                    return render(request, "helpdesk_hdincidente/editsolicitudespiezapartes.html", data)
                except Exception as ex:
                    pass

            elif action == 'solicitudespiezapartes':
                try:
                    data['title'] = u'Solicitudes Piezas y Partes'
                    search = None
                    if 's' in request.GET:
                        search = request.GET['s'].strip()
                        ss = search.split(' ')
                        if len(ss) == 1:
                            solicitudes = HdSolicitudesPiezaPartes.objects.filter(piezaparte__descripcion__icontains=search,status=True)
                        else:
                            solicitudes = HdSolicitudesPiezaPartes.objects.filter(piezaparte__descripcion__icontains=search,status=True)
                    else:
                        solicitudes = HdSolicitudesPiezaPartes.objects.filter(status=True)
                    paging = MiPaginador(solicitudes, 10)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['solicitudespiezapartes'] = page.object_list
                    return render(request, "helpdesk_hdincidente/solicitudespiezapartes.html", data)
                except Exception as ex:
                    pass

            elif action == 'addprioridad':
                try:
                    data['title'] = u'Adicionar Prioridad'
                    data['form'] = HdPrioridadForm()
                    return render(request, "helpdesk_hdincidente/addprioridad.html", data)
                except Exception as ex:
                    pass

            elif action=='delprioridad':
                try:
                    data['title'] = u'Eliminar Prioridad'
                    data['prioridad'] = HdPrioridad.objects.get(pk=int(request.GET['id']))
                    return render(request, "helpdesk_hdincidente/delprioridad.html", data)
                except Exception as ex:
                    pass

            elif action == 'editprioridad':
                try:
                    data['title'] = u'Editar Prioridad'
                    data['prioridad'] = prioridad=HdPrioridad.objects.get(pk=int(request.GET['id']))
                    form = HdPrioridadForm(initial={'nombre':prioridad.nombre,'codigo':prioridad.codigo, 'hora':prioridad.horamax, 'minuto':prioridad.minutomax, 'segundo':prioridad.segundomax})
                    if prioridad.esta_activo():
                        form.editar()
                    form.ocultarimagen()
                    data['form'] = form
                    return render(request, "helpdesk_hdincidente/editprioridad.html", data)
                except Exception as ex:
                    pass

            elif action == 'cambiarimagenprioridad':
                try:
                    data['title'] = u'Cambiar imagen de prioridad Help Desk'
                    data['prioridad'] = HdPrioridad.objects.get(pk=int(request.GET['id']))
                    form = HdPrioridadImagenForm()
                    data['form'] = form
                    return render(request, "helpdesk_hdincidente/cambiarimagenprioridad.html", data)
                except Exception as ex:
                    pass

            elif action == 'prioridad':
                try:
                    data['title'] = u'Lista de Prioridad'
                    data['prioridad'] = HdPrioridad.objects.filter(status=True).order_by('codigo')
                    return render(request, "helpdesk_hdincidente/prioridad.html", data)
                except Exception as ex:
                    pass
            #Medio de reporte
            elif action == 'addmedioreporte':
                try:
                    data['title'] = u'Adicionar Medio de reporte de Help Desk'
                    data['form'] = HdMedioReporteForm()
                    return render(request, "helpdesk_hdincidente/addmedioreporte.html", data)
                except Exception as ex:
                    pass

            elif action=='delmedioreporte':
                try:
                    data['title'] = u'Eliminar Medio de reporte de Help Desk'
                    data['medio'] = HdMedioReporte.objects.get(pk=int(request.GET['id']))
                    return render(request, "helpdesk_hdincidente/delmedioreporte.html", data)
                except Exception as ex:
                    pass

            elif action == 'editmedioreporte':
                try:
                    data['title'] = u'Editar Medio de reporte de Help Desk'
                    data['medio'] = medio = HdMedioReporte.objects.get(pk=int(request.GET['id']))
                    form = HdMedioReporteForm(initial={'nombre':medio.nombre, 'descripcion':medio.descripcion})
                    if medio.esta_activo():
                        form.editar()
                    data['form'] = form
                    return render(request, "helpdesk_hdincidente/editmedioreporte.html", data)
                except Exception as ex:
                    pass

            elif action == 'medioreporte':
                try:
                    data['title'] = u'Medio de reportes de Help Desk'
                    # if request.user.has_perm("sagest.puede_ingresar_configurar_helpdesk"):
                    data['medios'] = HdMedioReporte.objects.filter(status=True)
                    return render(request, "helpdesk_hdincidente/medioreporte.html", data)
                except Exception as ex:
                    pass
            #Prioridad Union
            elif action == 'addurgencia_impacto_prioridad':
                try:
                    data['title'] = u'Adicionar Prioridad'
                    form = HdUrgencia_Impacto_PrioridadForm()
                    data['form']=form
                    return render(request, "helpdesk_hdincidente/addunionimpactourgenciaprioridad.html", data)
                except Exception as ex:
                    pass

            elif action=='delurgencia_impacto_prioridad':
                try:
                    data['title'] = u'Eliminar Registro de Tabla de Prioridad'
                    data['gestion'] = HdUrgencia_Impacto_Prioridad.objects.get(pk=int(request.GET['id']))
                    return render(request, "helpdesk_hdincidente/delunionimpactourgenciaprioridad.html", data)
                except Exception as ex:
                    pass

            elif action == 'editurgencia_impacto_prioridad':
                try:
                    data['title'] = u'Editar Gestion de Prioridad'
                    data['gestion'] = gestion =HdUrgencia_Impacto_Prioridad.objects.get(pk=int(request.GET['id']))
                    if gestion.modificar:
                        form = HdUrgencia_Impacto_PrioridadForm(initial={'urgencia':gestion.urgencia,
                                                                         'impacto':gestion.impacto,
                                                                         'prioridad':gestion.prioridad,
                                                                         'modificar': gestion.modificar,
                                                                         'hora': gestion.horamax,
                                                                         'minuto': gestion.minutomax,
                                                                         'segundo': gestion.segundomax,
                                                                         })
                    else:
                        form = HdUrgencia_Impacto_PrioridadForm(initial={'urgencia': gestion.urgencia,
                                                                         'impacto': gestion.impacto,
                                                                         'prioridad': gestion.prioridad,
                                                                         'modificar': gestion.modificar,
                                                                         'hora': gestion.prioridad.horamax,
                                                                         'minuto': gestion.prioridad.minutomax,
                                                                         'segundo': gestion.prioridad.segundomax,
                                                                         })
                    data['form'] = form
                    return render(request, "helpdesk_hdincidente/editunionimpactourgenciaprioridad.html", data)
                except Exception as ex:
                    pass

            elif action == 'unionimpactourgenciaprioridad':
                try:
                    data['title'] = u'Tabla de Prioridad'
                    # if request.user.has_perm("sagest.puede_ingresar_configurar_helpdesk"):
                    search = None
                    ids = None
                    if 's' in request.GET:
                        search = request.GET['s']
                    if search:
                        grupo = HdUrgencia_Impacto_Prioridad.objects.filter(Q(prioridad__nombre__contains=search) | Q(urgencia__nombre__contains=search) | Q(impacto__nombre__contains=search), status=True)
                    elif 'id' in request.GET:
                        ids = request.GET['id']
                        grupo = HdUrgencia_Impacto_Prioridad.objects.filter(id=ids, status=True)
                    else:
                        grupo = HdUrgencia_Impacto_Prioridad.objects.all().exclude(status=False)
                    paging = MiPaginador(grupo, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['gestion'] = page.object_list
                    return render(request, "helpdesk_hdincidente/unionimpactourgenciaprioridad.html", data)
                except Exception as ex:
                    pass
            #Grupos de agentes
            elif action == 'addgrupo':
                try:
                    data['title'] = u'Adicionar Grupo'
                    data['form'] = HdGrupoForm()
                    return render(request, "helpdesk_hdincidente/addgrupo.html", data)
                except Exception as ex:
                    pass

            elif action == 'editgrupo':
                try:
                    data['title'] = u'Editar Grupo'
                    data['grupo'] = grupo = HdGrupo.objects.get(pk=int(request.GET['id']))
                    form = HdGrupoForm(initial={'grupo': grupo.nombre,'descripcion': grupo.descripcion, 'tipoincidente': grupo.tipoincidente})
                    data['form'] = form
                    return render(request, "helpdesk_hdincidente/editgrupo.html", data)
                except Exception as ex:
                    pass

            elif action == 'delgrupo':
                try:
                    data['title'] = u'Borrar Integrante'
                    data['grupo'] = HdGrupo.objects.get(pk=int(request.GET['id']))
                    return render(request, "helpdesk_hdincidente/delgrupo.html", data)
                except Exception as ex:
                    pass

            elif action == 'grupo':
                try:
                    data['title'] = u'Listado de Grupos Help Desk'
                    # if request.user.has_perm("sagest.puede_ingresar_configurar_helpdesk"):
                    search = None
                    ids = None
                    if 's' in request.GET:
                        search = request.GET['s']
                    if search:
                        grupo = HdGrupo.objects.filter(nombre__icontains=search, status=True)
                    elif 'id' in request.GET:
                        ids = request.GET['id']
                        grupo = HdGrupo.objects.filter(id=ids, status=True)
                    else:
                        grupo = HdGrupo.objects.all().exclude(status=False)
                    paging = MiPaginador(grupo, 10)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['grupo'] = page.object_list
                    return render(request, 'helpdesk_hdincidente/grupo.html', data)
                except Exception as ex:
                    pass
            #Agentes
            elif action == 'agentes':
                try:
                    # if request.user.has_perm("sagest.puede_ingresar_configurar_helpdesk"):
                    data['title'] = u'Agente Help Desk'
                    data['grupo'] = HdGrupo.objects.get(pk=int(request.GET['id']), status=True)
                    return render(request, "helpdesk_hdincidente/agente.html", data)
                except Exception as ex:
                    pass

            elif action == 'addagente':
                try:
                    data['title'] = u'Adicionar Agente'
                    data['grupo'] =grupo= HdGrupo.objects.get(pk=int(request.GET['idp']))
                    form = HdDetalle_GrupoForm()
                    if grupo.hddetalle_grupo_set.filter(responsable=True,estado=True).exists():
                        form.existe_responsable()
                    data['form'] = form
                    return render(request, "helpdesk_hdincidente/addagente.html", data)
                except Exception as ex:
                    pass

            elif action == 'delagente':
                try:
                    data['title'] = u'Borrar agente'
                    data['agente'] =grupo= HdDetalle_Grupo.objects.get(pk=int(request.GET['id']))
                    data['grupo'] = HdGrupo.objects.get(pk=grupo.grupo.id)
                    return render(request, "helpdesk_hdincidente/delagente.html", data)
                except Exception as ex:
                    pass
            #Categorias
            elif action == 'addcategoria':
                try:
                    data['title'] = u'Adicionar Categoria'
                    data['form'] = HdCatrgoriaForm()
                    return render(request, "helpdesk_hdincidente/addcategoria.html", data)
                except Exception as ex:
                    pass

            elif action=='delcategoria':
                try:
                    data['title'] = u'Eliminar Categoria'
                    data['categoria'] = HdCategoria.objects.get(pk=int(request.GET['id']))
                    return render(request, "helpdesk_hdincidente/delcategoria.html", data)
                except Exception as ex:
                    pass

            elif action == 'pdffechascortes':
                try:
                    data = {}
                    data['fechaactual'] = datetime.now()
                    data['nombrespersona'] = persona
                    grupousuario = persona.hddetalle_grupo_set.filter(status=True)[0].grupo
                    fechacierre = HdFechacierresolicitudes.objects.get(pk=request.GET['id'])
                    data['nombrespersonaresponsable'] = HdDetalle_Grupo.objects.filter(grupo=grupousuario, responsable=True, estado=True, status=True)[0].persona
                    # data['fechainicio'] = datetime.strptime(request.POST['ini'], "%d-%m-%Y").date()
                    # data['fechafin'] = datetime.strptime(request.POST['fin'], "%d-%m-%Y").date()
                    data['totalpiezaspartes'] = requerimientos = HdRequerimientosPiezaPartes.objects.values_list(
                        'solicitudes__piezaparte__descripcion', 'solicitudes__tipo', 'solicitudes__capacidad',
                        'solicitudes__velocidad', 'preciosolicitud__valor', 'solicitudes__descripcion').filter(Q(incidente__activo__archivobaja__isnull=True) | Q(incidente__activo__archivobaja=''),preciosolicitud__cierresolicitudes=fechacierre).annotate(sumuni=Sum('preciosolicitud__valor')).annotate(contador=Count('preciosolicitud__valor'))

                    # data['totalpiezaspartes'] = requerimientos = HdRequerimientosPiezaPartes.objects.values_list(
                    #     'solicitudes__piezaparte__descripcion', 'solicitudes__tipo', 'solicitudes__capacidad',
                    #     'solicitudes__velocidad', 'preciosolicitud__valor', 'solicitudes__descripcion').filter(
                    #     Q(incidente__activo__archivobaja__isnull=True) | Q(incidente__activo__archivobaja=''),
                    #     preciosolicitud__status=True, incidente__fechareporte__range=(
                    #     convertir_fecha(request.POST['ini']), convertir_fecha(request.POST['fin'])),
                    #     incidente__status=True, status=True).annotate(sumuni=Sum('preciosolicitud__valor')).annotate(contador=Count('preciosolicitud__valor'))

                    data['totalprecioreferencial'] = requerimientos.values_list('sumuni').aggregate(totalprecio=Sum('sumuni'))['totalprecio']
                    data['totalprecioreferencialunitario'] = requerimientos.values_list('preciosolicitud__valor').aggregate(totalsolicitudvalor=Sum('preciosolicitud__valor'))['totalsolicitudvalor']
                    data['totalcantidad'] = requerimientos.values_list('contador').aggregate(totalcontador=Sum('contador'))['totalcontador']
                    return conviert_html_to_pdf(
                        'helpdesk_hdincidente/pdflistapartes.html',
                        {
                            'pagesize': 'A4',
                            'data': data,
                        }
                    )
                except Exception as ex:
                    pass

            elif action == 'editcategoria':
                try:
                    data['title'] = u'Editar Categoria'
                    data['categoria'] = categoria = HdCategoria.objects.get(pk=int(request.GET['id']))
                    form = HdCatrgoriaForm(initial={'nombre':categoria,
                                                    'tipoincidente':categoria.tipoincidente})
                    if categoria.esta_activo():
                        form.editar()
                    data['form'] = form
                    return render(request, "helpdesk_hdincidente/editcategoria.html", data)
                except Exception as ex:
                    pass

            elif action == 'categoria':
                try:
                    # if request.user.has_perm("sagest.puede_ingresar_configurar_helpdesk"):
                    data['title'] = u'Lista de Categoría'
                    search = None
                    ids = None
                    if 's' in request.GET:
                        search = request.GET['s']
                    if search:
                        grupo = HdCategoria.objects.filter(nombre__icontains=search, status=True)
                    elif 'id' in request.GET:
                        ids = request.GET['id']
                        grupo = HdCategoria.objects.filter(id=ids, status=True)
                    else:
                        grupo = HdCategoria.objects.all().exclude(status=False)
                    paging = MiPaginador(grupo, 10)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['categoria'] = page.object_list
                    return render(request, "helpdesk_hdincidente/categoria.html", data)
                except Exception as ex:
                    pass
            #Tipo de incidentes
            elif action == 'addtipoincidente':
                try:
                    data['title'] = u'Adicionar tipo de incidente'
                    data['form'] = HdTipoIncidenteForm()
                    return render(request, "helpdesk_hdincidente/addtipoincidente.html", data)
                except Exception as ex:
                    pass

            elif action == 'addhdencuesta':
                try:
                    data['title'] = u'Adicionar encuesta'
                    data['idtipoincidente'] = HdTipoIncidente.objects.get(pk=int(request.GET['idtipoincidente']))
                    data['form'] = HdCabEncuestasForm()
                    return render(request, "helpdesk_hdincidente/addhdencuesta.html", data)
                except Exception as ex:
                    pass

            elif action == 'edithdencuesta':
                try:
                    data['title'] = u'Editar encuesta'
                    data['encuesta'] = encuesta = HdCabEncuestas.objects.get(pk=int(request.GET['idencuesta']))
                    form = HdCabEncuestasForm(initial={'activo':encuesta.activo, 'nombre':encuesta.nombre, 'descripcion': encuesta.descripcion})
                    data['form'] = form
                    return render(request, "helpdesk_hdincidente/edithdencuesta.html", data)
                except Exception as ex:
                    pass

            elif action == 'addhddetencuesta':
                try:
                    data['title'] = u'Adicionar Preguntas'
                    data['idencuesta'] = HdCabEncuestas.objects.get(pk=int(request.GET['idencuesta']))
                    data['form'] = HdDetEncuestasForm()
                    return render(request, "helpdesk_hdincidente/addhddetencuesta.html", data)
                except Exception as ex:
                    pass

            if action == 'editdetencuesta':
                try:
                    data['title'] = u'Editar Pregunta'
                    data['detencuesta'] = detencuesta = HdDetEncuestas.objects.get(pk=int(request.GET['iddetencuesta']))
                    form = HdDetEncuestasForm(initial={'activo':detencuesta.activo, 'pregunta':detencuesta.pregunta, 'tiporespuesta': detencuesta.tiporespuesta})
                    data['form'] = form
                    return render(request, "helpdesk_hdincidente/editdetencuesta.html", data)
                except Exception as ex:
                    pass

            elif action == 'deltipoincidente':
                try:
                    data['title'] = u'Eliminar tipo de inmcidente'
                    data['tipo'] = HdTipoIncidente.objects.get(pk=int(request.GET['id']))
                    return render(request, "helpdesk_hdincidente/deltipoincidente.html", data)
                except Exception as ex:
                    pass

            elif action == 'edittipoincidente':
                try:
                    data['title'] = u'Editar tipo de incidente'
                    data['tipo'] = tipo = HdTipoIncidente.objects.get(pk=int(request.GET['id']))
                    form = HdTipoIncidenteForm(initial={'nombre':tipo.nombre, 'descripcion': tipo.descripcion})
                    if tipo.esta_activo():
                        form.editar()
                    data['form'] = form
                    return render(request, "helpdesk_hdincidente/edittipoincidente.html", data)
                except Exception as ex:
                    pass

            elif action == 'listadoencuestas':
                try:
                    data['title'] = u'Listado de Encuestas'
                    search = None
                    ids = None
                    tipoincidente = HdTipoIncidente.objects.get(pk=int(request.GET['idtipo']))
                    if 's' in request.GET:
                        search = request.GET['s']
                    if search:
                        listadoencuestas = tipoincidente.tipoincidente.filter(nombre__icontains=search, status=True)
                    else:
                        listadoencuestas = tipoincidente.tipoincidente.filter(status=True).order_by('nombre')
                    paging = MiPaginador(listadoencuestas, 10)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['listadoencuestas'] = page.object_list
                    data['tipoincidente'] = tipoincidente
                    return render(request, "helpdesk_hdincidente/listadoencuestas.html", data)
                except Exception as ex:
                    pass

            elif action == 'listadopreguntas':
                try:
                    data['title'] = u'Listado de Preguntas'
                    search = None
                    ids = None
                    encuesta = HdCabEncuestas.objects.get(pk=int(request.GET['idencuesta']))
                    if 's' in request.GET:
                        search = request.GET['s']
                    if search:
                        listadopreguntas = encuesta.hdEncuestas.filter(status=True)
                    else:
                        listadopreguntas = encuesta.hdEncuestas.filter(status=True)
                    paging = MiPaginador(listadopreguntas, 10)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['listadopreguntas'] = page.object_list
                    data['encuesta'] = encuesta
                    return render(request, "helpdesk_hdincidente/listadopreguntas.html", data)
                except Exception as ex:
                    pass

            elif action == 'tipoincidente':
                try:
                    # if request.user.has_perm("sagest.puede_ingresar_configurar_helpdesk"):
                    data['title'] = u'Lista de tipos de incidentes'
                    search = None
                    ids = None
                    if 's' in request.GET:
                        search = request.GET['s']
                    if search:
                        tipos = HdTipoIncidente.objects.filter(nombre__icontains=search, status=True)
                    elif 'id' in request.GET:
                        ids = request.GET['id']
                        tipos = HdTipoIncidente.objects.filter(id=ids, status=True)
                    else:
                        tipos = HdTipoIncidente.objects.all().exclude(status=False)
                    paging = MiPaginador(tipos, 10)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['tipos'] = page.object_list
                    return render(request, "helpdesk_hdincidente/view_tipoincidentes.html", data)
                except Exception as ex:
                    pass
            #Sub Categoria
            elif action == 'addsubcategoria':
                try:
                    data['title'] = u'Adicionar Sub Categoria'
                    data['categoria'] = categoria = HdCategoria.objects.get(pk=int(request.GET['idc']))
                    form = HdSubCatrgoriaForm(initial={'categoria':categoria})
                    data['form'] = form
                    return render(request, "helpdesk_hdincidente/addsubcategoria.html", data)
                except Exception as ex:
                    pass

            elif action == 'editsubcategoria':
                try:
                    data['title'] = u'Editar Sub Categoria'
                    data['categoria'] = categoria = HdCategoria.objects.get(pk=int(request.GET['id']))
                    data['subcat'] = subcat = HdSubCategoria.objects.get(pk=int(request.GET['idsubc']))
                    form = HdSubCatrgoriaForm(initial={'categoria':categoria, 'subcategoria':subcat.nombre})
                    if categoria.esta_activo():
                        form.editar()
                    data['form'] = form
                    return render(request, "helpdesk_hdincidente/editsubcategoria.html", data)
                except Exception as ex:
                    pass

            elif action=='delsubcategoria':
                try:
                    data['title'] = u'Eliminar Sub Categoria'
                    data['categoria'] = HdCategoria.objects.get(pk=int(request.GET['id']))
                    data['subcat'] = HdSubCategoria.objects.get(pk=int(request.GET['idsubc']))
                    return render(request, "helpdesk_hdincidente/delsubcategoria.html", data)
                except Exception as ex:
                    pass

            elif action == 'subcategoria':
                try:
                    data['title'] = u'Adicionar SubCategoria'
                    data['categoria']= cat =HdCategoria.objects.get(pk=int(request.GET['idc']))
                    search = None
                    ids = None
                    if 's' in request.GET:
                        search = request.GET['s']
                    if search:
                        categoria = cat.hdsubcategoria_set.filter(nombre__icontains=search, status=True)
                    elif 'id' in request.GET:
                        ids = request.GET['id']
                        categoria = HdSubCategoria.objects.filter(id=ids, status=True)
                    else:
                        categoria = cat.hdsubcategoria_set.all().exclude(status=False)
                    paging = MiPaginador(categoria, 10)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['categorias'] = page.object_list
                    return render(request, "helpdesk_hdincidente/subcategoria.html", data)
                except Exception as ex:
                    pass
            #Detalle de subcategoria
            elif action == 'adddetalle':
                try:
                    data['title'] = u'Adicionar Detalle'
                    data['subcat'] = subcat = HdSubCategoria.objects.get(pk=int(request.GET['idsubc']))
                    form = HdDetalleSubCategoriaForm(initial={'subcategoria':subcat.nombre})
                    form.editar()
                    data['form'] = form
                    return render(request, "helpdesk_hdincidente/adddetalle.html", data)
                except Exception as ex:
                    pass

            elif action == 'aditdatallesubcategoria':
                try:
                    data['title'] = u'Editar Detalle de Sub categoria'
                    data['detalle'] = detalle = HdDetalle_SubCategoria.objects.get(pk=int(request.GET['id']))
                    data['subcat'] = subcat = HdSubCategoria.objects.get(pk=int(request.GET['idsubc']))
                    # tiempo = ''
                    # if detalle.prioridad.modificar and int(detalle.prioridad.horamax) > 0:
                    if detalle.prioridad:
                        if detalle.prioridad.modificar and int(detalle.prioridad.horamax) > 0:
                            tiempo = detalle.prioridad.horamax + ":" + detalle.prioridad.minutomax + ":" + detalle.prioridad.segundomax
                        else:
                            tiempo = detalle.prioridad.prioridad.horamax + ":" + detalle.prioridad.prioridad.minutomax + ":" + detalle.prioridad.prioridad.segundomax
                        form = HdDetalleSubCategoriaForm(initial={'subcategoria':subcat.nombre, 'detalle': detalle.nombre, 'urgencia': detalle.prioridad.urgencia, 'impacto': detalle.prioridad.impacto, 'tiemporesolucion': tiempo, 'prioridad':detalle.prioridad})
                    else:
                        form = HdDetalleSubCategoriaForm(initial={'subcategoria':subcat.nombre, 'detalle': detalle.nombre })
                    form.editar()
                    data['form'] = form
                    return render(request, "helpdesk_hdincidente/editdetallesubcategoria.html", data)
                except Exception as ex:
                    pass

            elif action=='deldetalle':
                try:
                    data['title'] = u'Eliminar  Detalle Sub Categoria'
                    data['detalle'] = HdDetalle_SubCategoria.objects.get(pk=int(request.GET['id']))
                    data['subcat'] = HdSubCategoria.objects.get(pk=int(request.GET['idsubc']))
                    return render(request, "helpdesk_hdincidente/deldetalle.html", data)
                except Exception as ex:
                    pass

            elif action == 'detalle':
                try:
                    data['title'] = u'Detalle de Sub Categoria'
                    data['subcategoria']=subcat = HdSubCategoria.objects.get(pk=int(request.GET['idsubc']))
                    search = None
                    ids = None
                    if 's' in request.GET:
                        search = request.GET['s']
                    if search:
                        detalle = subcat.hddetalle_subcategoria_set.filter(nombre__icontains=search, status=True)
                    elif 'id' in request.GET:
                        ids = request.GET['id']
                        detalle = HdDetalle_SubCategoria.objects.filter(id=ids, status=True)
                    else:
                        detalle = subcat.hddetalle_subcategoria_set.all().exclude(status=False)
                    paging = MiPaginador(detalle, 10)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['subcat'] = page.object_list
                    return render(request, "helpdesk_hdincidente/detalle.html", data)
                except Exception as ex:
                    pass
            #Proceso de baja
            elif action == 'addproceso':
                try:
                    data['title'] = u'Adicionar nuevo proceso'
                    data['form'] = HdProcesoFrom()
                    return render(request, "helpdesk_hdincidente/addproceso.html", data)
                except Exception as ex:
                    pass

            elif action == 'editproceso':
                try:
                    data['title'] = u'Editar Proceso'
                    data['proceso'] = proceso = HdProceso.objects.get(pk=int(request.GET['id']))
                    data['form'] = HdProcesoFrom(initial={'nombre':proceso.nombre})
                    return render(request, "helpdesk_hdincidente/editproceso.html", data)
                except Exception as ex:
                    pass

            elif action == 'delproceso':
                try:
                    data['title'] = u'Eliminar Proceso'
                    data['proceso'] = HdProceso.objects.get(pk=int(request.GET['id']))
                    return render(request, "helpdesk_hdincidente/delproceso.html", data)
                except Exception as ex:
                    pass

            elif action == 'viewproceso':
                try:
                    # if request.user.has_perm("sagest.puede_ingresar_configurar_helpdesk"):
                    data['title'] = u'Procesos de Incidente de Help Desk'
                    data['procesos'] = HdProceso.objects.all()
                    return render(request, "helpdesk_hdincidente/viewproceso.html", data)
                except Exception as ex:
                    pass
            #Estado  de Proceso de baja
            elif action == 'addestado_proceso':
                try:
                    data['title'] = u'Adicionar nuevo estado'
                    data['proceso'] =proceso= HdProceso.objects.get(pk=int(request.GET['id']))
                    data['form'] = HdEstadoProcesoFrom(initial={'proceso':proceso})
                    return render(request, "helpdesk_hdincidente/addestado_proceso.html", data)
                except Exception as ex:
                    pass

            elif action == 'editestado_proceso':
                try:
                    data['title'] = u'Editar Estado Proceso'
                    data['estadopro'] = estadopro = HdEstado_Proceso.objects.get(pk=int(request.GET['id']))
                    form = HdEstadoProcesoFrom(initial={'nombre':estadopro.nombre, 'proceso':estadopro.proceso,'detalle':estadopro.detalle})
                    data['form'] = form
                    return render(request, "helpdesk_hdincidente/editestado_proceso.html", data)
                except Exception as ex:
                    pass

            elif action == 'delestado_proceso':
                try:
                    data['title'] = u'Eliminar Estado Proceso'
                    data['estadopro'] = HdEstado_Proceso.objects.get(pk=int(request.GET['id']))
                    return render(request, "helpdesk_hdincidente/delestado_proceso.html", data)
                except Exception as ex:
                    pass

            elif action == 'addagregarestado_proceso':
                try:
                    data['title'] = u'Adicionar nuevo estado de proceso'
                    data['proceso'] = proceso = HdProceso.objects.get(pk=int(request.GET['id']))
                    data['form'] = HdEstadoProcesoFrom(initial={'proceso':proceso})
                    return render(request, "helpdesk_hdincidente/addagregarestado_proceso.html", data)
                except Exception as ex:
                    pass

            elif action == 'viewestado_proceso':
                try:
                    data['title'] = u'Estado de Procesos de Incidente'
                    # if request.user.has_perm("sagest.puede_ingresar_configurar_helpdesk"):
                    if 'id' in request.GET:
                        data['proceso']= HdProceso.objects.get(pk=int(request.GET['id']))
                        data['estadospro'] = HdEstado_Proceso.objects.filter(proceso_id=int(request.GET['id']))
                        return render(request, "helpdesk_hdincidente/viewestado_proceso.html", data)
                except Exception as ex:
                    pass
            #Director
            elif action == 'adddirector':
                try:
                    data['title'] = u'Adicionar nuevo Director'
                    data['form'] = HdDirectorForm()
                    # if HdDirector.objects.filter(status=True).exists():
                    #     data['existe']=True
                    return render(request, "helpdesk_hdincidente/adddirector.html", data)
                except Exception as ex:
                    pass

            elif action == 'editdirector':
                try:
                    data['title'] = u'Editar Director'
                    data['director'] = director = HdDirector.objects.get(pk=int(request.GET['id']))
                    form = HdDirectorForm(initial={'director':director.persona.nombre_completo()})
                    data['form'] = form
                    return render(request, "helpdesk_hdincidente/editdirector.html", data)
                except Exception as ex:
                    pass

            elif action == 'deldirector':
                try:
                    data['title'] = u'Eliminar Director'
                    data['director'] = HdDirector.objects.get(pk=int(request.GET['id']))
                    return render(request, "helpdesk_hdincidente/deldirector.html", data)
                except Exception as ex:
                    pass

            elif action == 'viewdirector':
                try:
                    data['title'] = u'Director del Departamento'
                    # if request.user.has_perm("sagest.puede_ingresar_configurar_helpdesk"):
                    data['director']= HdDirector.objects.filter(status=True).order_by('persona')
                    return render(request, "helpdesk_hdincidente/viewdirector.html", data)
                except Exception as ex:
                    pass

            elif action == 'bloque_ubicacion':
                try:
                    # if request.user.has_perm("sagest.puede_ingresar_configurar_helpdesk"):
                    data['title'] = u'Bloques del campus Universitario'
                    data['bloqueubicaciones'] = HdBloqueUbicacion.objects.filter(status=True).distinct('bloque').order_by('bloque')
                    return render(request, "helpdesk_hdincidente/view_bloque_ubicacion.html", data)
                except Exception as ex:
                    pass

            elif action == 'listar_bloque':
                try:
                    lista_bloque=[]
                    lista_ubicaciones = []
                    for blo in HdBloque.objects.filter(status=True).order_by('id'):
                        lista_bloque.append([blo.id, str(blo.nombre)])
                    for ubi in HdUbicacion.objects.filter(status=True).order_by('id'):
                        lista_ubicaciones.append([ubi.id, str(ubi.nombre)])
                    data = {"result": "ok", 'lista_bloque': lista_bloque, 'lista_ubicaciones': lista_ubicaciones}
                    return JsonResponse(data)
                except Exception as ex:
                    pass

            elif action == 'no_existe_registro':
                try:
                    if 'idb' in request.GET and 'idu' in request.GET:
                        if not HdBloqueUbicacion.objects.filter(status=True, bloque_id=request.GET['idb'], ubicacion_id=request.GET['idu']).exists():
                            data = {"result": "ok"}
                        else:
                            data = {"result": "none"}
                        return JsonResponse(data)
                except Exception as ex:
                    pass

            elif action == 'delbloque_ubicacion':
                try:
                    data['title'] = u'Eliminar ubicación'
                    data['bloque'] = HdBloqueUbicacion.objects.get(pk=int(request.GET['id']))
                    return render(request, "helpdesk_hdincidente/delbloque_ubicacion.html", data)
                except Exception as ex:
                    pass

            elif action == 'delgrupobloque_ubicacion':
                try:
                    data['title'] = u'Eliminar bloque y ubicaciones'
                    data['bloque'] = HdBloqueUbicacion.objects.get(pk=int(request.GET['id']))
                    return render(request, "helpdesk_hdincidente/delgrupobloque_ubicacion.html", data)
                except Exception as ex:
                    pass

            elif action == 'bloque':
                try:
                    # if request.user.has_perm("sagest.puede_ingresar_configurar_helpdesk"):

                    data['title'] = u'Lista de bloques del campus Universitario'
                    search = None
                    ids = None
                    if 's' in request.GET:
                        search = request.GET['s']
                        search = request.GET['s'].strip()
                        ss = search.split(' ')
                        if len(ss)==1:
                            bloques = HdBloque.objects.filter(nombre__icontains=search, status=True)
                        else:
                            bloques = HdBloque.objects.filter(Q(nombre__icontains=ss[0]), Q(nombre__icontains=ss[1]), status=True)
                    elif 'id' in request.GET:
                        ids = request.GET['id']
                        bloques = HdBloque.objects.filter(id=ids, status=True)
                    else:
                        bloques = HdBloque.objects.filter(status=True)
                    paging = MiPaginador(bloques, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['bloques'] = page.object_list
                    # data['bloques'] = HdBloque.objects.filter(status=True).order_by('id')
                    return render(request, "helpdesk_hdincidente/view_bloque.html", data)
                except Exception as ex:
                    pass

            elif action == 'editbloque':
                try:
                    data['title'] = u'Editar Bloque'
                    ubi = HdBloque.objects.get(pk=int(request.GET['id']))
                    data = {"result": "ok", 'ubicacion': ubi.nombre, 'id': ubi.id}
                    return JsonResponse(data)
                except Exception as ex:
                    pass

            elif action == 'delbloque':
                try:
                    data['title'] = u'Eliminar Bloque'
                    data['blo'] = blo = HdBloque.objects.get(pk=int(request.GET['id']))
                    return render(request, "helpdesk_hdincidente/delbloque.html", data)
                except Exception as ex:
                    pass

            elif action == 'ubicacion':
                try:
                    # if request.user.has_perm("sagest.puede_ingresar_configurar_helpdesk"):
                    data['title'] = u'Ubicación del campus Universitario'
                    search = None
                    ids = None
                    if 's' in request.GET:
                        search = request.GET['s']
                        search = request.GET['s'].strip()
                        ss = search.split(' ')
                        if len(ss) == 1:
                            ubicaciones = HdUbicacion.objects.filter(nombre__icontains=search, status=True)
                        else:
                            ubicaciones = HdUbicacion.objects.filter(Q(nombre__icontains=ss[0]), Q(nombre__icontains=ss[1]),status=True)
                    elif 'id' in request.GET:
                        ids = request.GET['id']
                        ubicaciones = HdUbicacion.objects.filter(id=ids, status=True)
                    else:
                        ubicaciones = HdUbicacion.objects.filter(status=True)
                    paging = MiPaginador(ubicaciones, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['ubicaciones'] = page.object_list
                    # data['ubicaciones'] = HdUbicacion.objects.filter(status=True).order_by('id')
                    return render(request, "helpdesk_hdincidente/view_ubicacion.html", data)
                except Exception as ex:
                    pass

            elif action == 'editubicacion':
                try:
                    data['title'] = u'Editar Ubicacion'
                    ubi = HdUbicacion.objects.get(pk=int(request.GET['id']))
                    data = {"result": "ok", 'ubicacion': ubi.nombre, 'id': ubi.id}
                    return JsonResponse(data)
                except Exception as ex:
                    pass

            elif action == 'delubicacion':
                try:
                    data['title'] = u'Eliminar Ubicación'
                    data['ubi'] = HdUbicacion.objects.get(pk=int(request.GET['id']))
                    return render(request, "helpdesk_hdincidente/delubicacion.html", data)
                except Exception as ex:
                    pass

            elif action == 'selectubicacion':
                try:
                    if 'id' in request.GET:
                        lista = []
                        ubicaciones = HdBloqueUbicacion.objects.filter(bloque_id=int(request.GET['id']),status=True)
                        for ubi in ubicaciones:
                            lista.append([ubi.id, ubi.ubicacion.nombre])
                        data = {"results": "ok", 'lista': lista}
                        return JsonResponse(data)
                except Exception as ex:
                    pass

            elif action == 'reportegeneral_excel':
                try:
                    __author__ = 'Unemi'
                    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
                    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                    title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    style1 = easyxf(num_format_str='D-MMM-YY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('exp_xls_post_part')
                    ws.write_merge(0, 0, 0, 7, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    ws.write_merge(1, 1, 0, 2, 'De ' +request.GET['de']+' hasta '+ request.GET['hasta'], style_sb)
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=Incidentes ' + random.randint(1, 10000).__str__() + '.xls'
                    columns = [
                        (u"FECHA REPORTE", 4500),
                        (u"ASUNTO", 15000),
                        (u"ACTIVO", 15000),
                        (u"ACTIVO CODIGO INTERNO", 10000),
                        (u"ACTIVO CODIGO DE GOBIERNO", 10000),
                        (u"AGENTE", 11000),
                        (u"GRUPO", 10000),
                        (u"CATEGORÍA", 13000),
                        (u"SUB CATEGORÍA", 13000),
                        (u"DETALLE DE SUB CATEGORÍA", 15000),
                        (u"CAUSAS", 15000),
                        (u"RESOLUCIÓN", 15000),
                        (u"ORDEN DE TRABAJO", 1500),
                        (u"ESTADO", 1500),
                    ]
                    row_num = 3
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    date_format = xlwt.XFStyle()
                    date_format.num_format_str = 'yyyy/mm/dd'
                    incidentes = HdIncidente.objects.filter(status=True, fechareporte__range=(convertir_fecha(request.GET['de']), convertir_fecha(request.GET['hasta']))).order_by('fechareporte')
                    row_num = 4
                    for incidente in incidentes:
                        ws.write(row_num, 0, incidente.fechareporte, date_format)
                        ws.write(row_num, 1, incidente.asunto if incidente.asunto else '', font_style2)
                        ws.write(row_num, 2, "%s  %s" % (incidente.activo.descripcion, incidente.activo.modelo) if incidente.activo else '', font_style2)
                        ws.write(row_num, 3, "%s" % incidente.activo.codigointerno if incidente.activo else '', font_style2)
                        ws.write(row_num, 4, "%s" % incidente.activo.codigogobierno if incidente.activo else '', font_style2)
                        ws.write(row_num, 5, incidente.ultimo_agente_asignado().persona.nombre_completo_inverso() if incidente.ultimo_agente_asignado() else '', font_style2)
                        ws.write(row_num, 6, incidente.mi_detalle()[0].grupo.nombre if incidente.mi_detalle() else '', font_style2)
                        ws.write(row_num, 7, incidente.detallesubcategoria.subcategoria.categoria.nombre if incidente.detallesubcategoria else '', font_style2)
                        ws.write(row_num, 8, incidente.detallesubcategoria.subcategoria.nombre if incidente.detallesubcategoria else '', font_style2)
                        ws.write(row_num, 9, incidente.detallesubcategoria.nombre if incidente.detallesubcategoria else '', font_style2)
                        ws.write(row_num, 10, incidente.causa.nombre if incidente.causa else '', font_style2)
                        ws.write(row_num, 11, incidente.mi_detalle()[0].resolucion if incidente.mi_detalle() else '', font_style2)
                        ws.write(row_num, 12,"%s" % incidente.ordentrabajo.codigoorden if incidente.ordentrabajo_id else '', font_style2)
                        ws.write(row_num, 13,"%s" % incidente.estado, font_style2)
                        row_num += 1
                    # incidentesmem = incidentes.values_list('hdrequerimientospiezapartes__solicitudes__piezaparte__descripcion').filter(status=True).distinct()
                    columns = [
                        (u"FECHA REPORTE", 4500),
                        (u"ASUNTO", 15000),
                        (u"AGENTE", 11000),
                        (u"GRUPO", 10000),
                        (u"CATEGORÍA", 13000),
                        (u"SUB CATEGORÍA", 13000),
                        (u"DETALLE DE SUB CATEGORÍA", 15000),
                        (u"Causa", 15000),
                        (u"RESOLUCIÓN", 15000),
                        (u"TIPO", 3000),
                        (u"CAPACIDAD", 3000),
                        (u"VELOCIDAD", 3000),
                    ]
                    incidentesmem = HdPiezaPartes.objects.filter(status=True)
                    for parte in incidentesmem:
                        row_num = 0
                        if incidentes.filter(hd_incidentes__solicitudes__piezaparte__id=parte.id, status=True).exists():
                            incidentesmem = incidentes.filter(hd_incidentes__solicitudes__piezaparte__id=parte.id, status=True).distinct()
                            wmem = wb.add_sheet(parte.descripcion)
                            for col_num in range(len(columns)):
                                wmem.write(row_num, col_num, columns[col_num][0], font_style)
                                wmem.col(col_num).width = columns[col_num][1]
                            row_num = 1
                            for listaincidentes in incidentesmem:
                                wmem.write(row_num, 0, listaincidentes.fechareporte, date_format)
                                wmem.write(row_num, 1, listaincidentes.asunto if listaincidentes.asunto else '', font_style2)
                                wmem.write(row_num, 2, listaincidentes.ultimo_agente_asignado().persona.nombre_completo_inverso() if listaincidentes.ultimo_agente_asignado() else '',font_style2)
                                wmem.write(row_num, 3,listaincidentes.mi_detalle()[0].grupo.nombre if listaincidentes.mi_detalle() else '', font_style2)
                                wmem.write(row_num, 4,listaincidentes.detallesubcategoria.subcategoria.categoria.nombre if listaincidentes.detallesubcategoria else '', font_style2)
                                wmem.write(row_num, 5,listaincidentes.detallesubcategoria.subcategoria.nombre if listaincidentes.detallesubcategoria else '', font_style2)
                                wmem.write(row_num, 6,listaincidentes.detallesubcategoria.nombre if listaincidentes.detallesubcategoria else '',  font_style2)
                                wmem.write(row_num, 7, listaincidentes.causa.nombre if listaincidentes.causa else '', font_style2)
                                wmem.write(row_num, 8, listaincidentes.mi_detalle()[0].resolucion if listaincidentes.mi_detalle() else '', font_style2)
                                wmem.write(row_num, 9, listaincidentes.hd_incidentes.values_list('solicitudes__tipo').filter(status=True)[0] if listaincidentes.hdrequerimientospiezapartes_set.values_list('solicitudes__tipo').filter(status=True)[0] else '', font_style2)
                                wmem.write(row_num, 10, listaincidentes.hd_incidentes.values_list('solicitudes__capacidad').filter(status=True)[0] if listaincidentes.hdrequerimientospiezapartes_set.values_list('solicitudes__capacidad').filter(status=True)[0] else '', font_style2)
                                wmem.write(row_num, 11, listaincidentes.hd_incidentes.values_list('solicitudes__velocidad').filter(status=True)[0] if listaincidentes.hdrequerimientospiezapartes_set.values_list('solicitudes__velocidad').filter(status=True)[0] else '', font_style2)
                                row_num += 1

                    wsb = wb.add_sheet('listado_bajas')
                    wsb.write_merge(0, 0, 0, 7, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    wsb.write_merge(1, 1, 0, 2, 'De ' + request.GET['de'] + ' hasta ' + request.GET['hasta'], style_sb)
                    columns = [
                        (u"CODIGO GOBIERNO", 4500),
                        (u"CODIGO INTERNO", 4500),
                        (u"RESPONSABLE", 15000),
                        (u"BLOQUE", 4000),
                        (u"MAL USO", 4000),
                        (u"ESTADO", 4000),
                        (u"AGENTE", 15000),
                        (u"DESCRIPCION", 15000),
                    ]
                    row_num = 3
                    for col_num in range(len(columns)):
                        wsb.write(row_num, col_num, columns[col_num][0], font_style)
                        wsb.col(col_num).width = columns[col_num][1]
                    listadoinformebaja = InformeActivoBaja.objects.filter(fecha_creacion__range=(convertir_fecha(request.GET['de']), convertir_fecha(request.GET['hasta'])), status=True)
                    row_num = 4
                    for activobaja in listadoinformebaja:
                        if activobaja.activofijo.responsable:
                            nomresponsable = activobaja.activofijo.responsable.apellido1 + ' ' + activobaja.activofijo.responsable.apellido2 + ' ' + activobaja.activofijo.responsable.nombres
                        else:
                            nomresponsable = ''
                        person = Persona.objects.get(usuario=activobaja.usuario_creacion)
                        agente = person.apellido1 + ' ' + person.apellido2 + ' ' + person.nombres
                        wsb.write(row_num, 0, activobaja.activofijo.codigogobierno, font_style2)
                        wsb.write(row_num, 1, activobaja.activofijo.codigointerno, font_style2)
                        wsb.write(row_num, 2, nomresponsable,font_style2)
                        wsb.write(row_num, 3, activobaja.bloque.nombre,font_style2)
                        wsb.write(row_num, 4, activobaja.get_estadouso_display(),font_style2)
                        wsb.write(row_num, 5, activobaja.get_estado_display(),font_style2)
                        wsb.write(row_num, 6, agente,font_style2)
                        wsb.write(row_num, 7, activobaja.activofijo.descripcion,font_style2)
                        row_num += 1
                    wb.save(response)
                    return response
                except Exception as ex:
                    pass
            elif action == 'reportegeneral_excelcompleto':
                try:
                    tecnico=None
                    fechadesde = request.GET['de']
                    fechahasta = request.GET['hasta']
                    tec = int(request.GET['tecnico'])
                    bloque = int(request.GET['bloque'])
                    estado = int(request.GET['estado'])
                    checktodos = int(request.GET['checktodos'])
                    if checktodos == 1:
                        tec = 0
                    if not tec == 0:
                        tecnico = Administrativo.objects.get(pk=int(request.GET['tecnico'])).persona.pk
                        data['tecnico'] = Persona.objects.get(pk=int(tecnico)).nombre_completo_inverso()
                    else:
                        data['tecnico'] = 'TODOS LOS TECNICOS'
                    if not bloque == 0:
                        data['bloque'] = HdBloque.objects.get(pk=int(bloque)).nombre
                    else:
                        data['bloque'] = 'TODOS LOS BLOQUES'

                    esta = ''
                    if not estado == 0:
                        if estado == '1':
                            esta = 'GENERADO'
                        elif estado == '2':
                            esta = 'CERRADO'
                        elif estado == '3':
                            esta = 'PENDIENTE REPUESTO'
                        elif estado == '4':
                            esta = 'TALLER PARTICULAR'
                        elif estado == '5':
                            esta = 'EN TRÁMITE'
                        data['estado'] = esta
                    else:
                        data['estado'] = 'TODOS LOS ESTADOS'

                    __author__ = 'Unemi'
                    borders = Borders()
                    borders.left = 1
                    borders.right = 1
                    borders.top = 1
                    borders.bottom = 1
                    styrow = easyxf( 'font: name Times New Roman, color-index black, bold off; alignment: horiz centre; borders: left thin, right thin, top thin, bottom thin')
                    style_col = easyxf( 'pattern: pattern solid, fore_colour light_turquoise; font: name Times New Roman, color-index black, bold off; alignment: horiz centre; borders: left thin, right thin, top thin, bottom thin')
                    style_sb1 = easyxf('font: name Times New Roman, color-index black, bold on')
                    style_sb = easyxf('font: name Times New Roman, color-index black, bold off')
                    title = easyxf( 'font: name Times New Roman, color-index green, bold on , height 350; alignment: horiz centre')
                    title1 = easyxf( 'font: name Times New Roman, color-index green, bold on , height 250; alignment: horiz centre')
                    style_date = easyxf( 'borders: left thin, right thin, top thin, bottom thin; alignment: horiz centre',num_format_str='yy/mm/dd h:mm')


                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('exp_xls_post_part')
                    # ti1=('TÉCNICO RESPONSABLE: ',style_sb)

                    ws.write_merge(0, 0, 0, 5, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    ws.write_merge(1, 1, 0, 5, 'ÁREA DE MANTENIMIENTO', title1)
                    ws.write_merge(2, 2, 0, 5, 'REPORTE RESUMEN DE ÓRDENES DE TRABAJO', title1)
                    ws.write_merge(4, 4, 0, 0, 'PERIODO:  ', style_sb1)
                    ws.write_merge(4, 4, 1, 1, 'DESDE  ' + request.GET['de'] + ' HASTA  ' + request.GET['hasta'],
                                   style_sb)
                    ws.write_merge(5, 5, 0, 0, 'TÉCNICO RESPONSABLE: ', style_sb1)
                    ws.write_merge(5, 5, 1, 1, data['tecnico'], style_sb)
                    ws.write_merge(6, 6, 0, 0, 'BLOQUE: ', style_sb1)
                    ws.write_merge(6, 6, 1, 1, data['bloque'], style_sb)
                    ws.write_merge(7, 7, 0, 0, 'ESTADO: ', style_sb1)
                    ws.write_merge(7, 7, 1, 1, data['estado'], style_sb)
                    # ws.write_merge(3, 3, 0, 2, 'De ' +request.GET['de']+' hasta '+ request.GET['hasta'], style_sb)
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=ResumenOrdenes ' + random.randint(1,10000).__str__() + '.xls'

                    columns = [
                        (u"N° ORDEN", 5800),
                        (u"FECHA REGISTRO", 9000),
                        (u"ESTADO", 4500),
                        (u"FECHA EJECUCIÓN MANTENIMIENTO", 10000),
                        (u"FECHA DE CIERRE", 10000),
                        (u"TÉCNICO RESPONSABLE", 11000),

                    ]
                    row_num = 9
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], style_col)
                        ws.col(col_num).width = columns[col_num][1]

                    #incident=""
                    #if tec == 0 and bloque == 0 and estado == 0:
                        #incidentes = HdDetalle_Incidente.objects.filter(
                    #incident =
                    # incidentes = HdDetalle_Incidente.objects.filter(
                    #             incidente__ordentrabajo__fecha_creacion__range=(
                    #                 convertir_fecha(request.GET['de']), convertir_fecha(request.GET['hasta'])),
                    #             incidente__status=True,
                    #             agente__persona=int(tecnico),
                    #             incidente__ubicacion__bloque=int(bloque),
                    #             incidente__ordentrabajo__estado=int(estado))

                    # incidentes = HdDetalle_Incidente.objects.filter(status=True)
                    # if not tec == 0:
                    #     incidentes = incidentes.filter(agente__persona=int(tecnico))
                    # if not bloque == 0:
                    #     incidentes = incidentes.filter(incidente__ubicacion__bloque=int(bloque))
                    # if not estado == 0:
                    #     incidentes = incidentes.filter(incidente__ordentrabajo__estado=int(estado))


                    if bloque == 0 and not estado == 0 and not tec == 0:
                        incidentes = HdDetalle_Incidente.objects.filter(incidente__ordentrabajo__fecha_creacion__range=(
                            convertir_fecha(request.GET['de']), convertir_fecha(request.GET['hasta'])),
                            agente__persona=int(tecnico), incidente__ordentrabajo__estado=int(estado),
                            incidente__status=True)
                    elif estado == 0 and not bloque == 0 and not tec == 0:
                        incidentes = HdDetalle_Incidente.objects.filter(incidente__ordentrabajo__fecha_creacion__range=(
                            convertir_fecha(request.GET['de']), convertir_fecha(request.GET['hasta'])),
                            incidente__status=True, agente__persona=int(tecnico),
                            incidente__ubicacion__bloque=int(bloque))
                    elif tec == 0 and not estado == 0 and not bloque == 0:
                        incidentes = HdDetalle_Incidente.objects.filter(incidente__ordentrabajo__fecha_creacion__range=(
                            convertir_fecha(request.GET['de']), convertir_fecha(request.GET['hasta'])),
                            incidente__status=True,
                            incidente__ubicacion__bloque=int(bloque), incidente__ordentrabajo__estado=int(estado))
                    elif tec == 0 and estado == 0 and bloque == 0:
                        incidentes = HdDetalle_Incidente.objects.filter(incidente__ordentrabajo__fecha_creacion__range=(
                            convertir_fecha(request.GET['de']), convertir_fecha(request.GET['hasta'])),
                            incidente__status=True )
                    elif estado == 0 and bloque == 0 and not tec == 0:
                        incidentes = HdDetalle_Incidente.objects.filter(incidente__ordentrabajo__fecha_creacion__range=(
                            convertir_fecha(request.GET['de']), convertir_fecha(request.GET['hasta'])),
                            incidente__status=True,
                            agente__persona=int(tecnico))
                    elif estado == 0 and tec == 0 and not bloque == 0:
                        incidentes = HdDetalle_Incidente.objects.filter(incidente__ordentrabajo__fecha_creacion__range=(
                            convertir_fecha(request.GET['de']), convertir_fecha(request.GET['hasta'])),
                            incidente__status=True,
                            incidente__ubicacion__bloque=int(bloque))
                    elif tec == 0 and bloque == 0 and not estado == 0:
                        incidentes = HdDetalle_Incidente.objects.filter(incidente__ordentrabajo__fecha_creacion__range=(
                            convertir_fecha(request.GET['de']), convertir_fecha(request.GET['hasta'])),
                            incidente__status=True,
                            incidente__ordentrabajo__estado=int(estado))
                    else:
                        incidentes = HdDetalle_Incidente.objects.filter(
                            incidente__ordentrabajo__fecha_creacion__range=(
                                convertir_fecha(request.GET['de']), convertir_fecha(request.GET['hasta'])),
                            incidente__status=True,
                            agente__persona=int(tecnico),
                            incidente__ubicacion__bloque=int(bloque),
                            incidente__ordentrabajo__estado=int(estado))

                    row_num = 10
                    for incidente in incidentes:
                        estado = ''
                        if incidente.incidente.ordentrabajo.estado == 1:
                            estado = 'GENERADO'
                        elif incidente.incidente.ordentrabajo.estado == 2:
                            estado = 'CERRADO'
                        elif incidente.incidente.ordentrabajo.estado == 3:
                            estado = 'PENDIENTE REPUESTO'
                        elif incidente.incidente.ordentrabajo.estado == 4:
                            estado = 'TALLER PARTICULAR'
                        elif incidente.ordentrabajo.estado == 5:
                            estado = 'EN TRÁMITE'
                        ws.write(row_num, 0, incidente.incidente.ordentrabajo.codigoorden, styrow)
                        ws.write(row_num, 1, (incidente.incidente.fecha_creacion), style_date)
                        ws.write(row_num, 2, str(estado), styrow)
                        ws.write(row_num, 3, (incidente.incidente.ordentrabajo.fecha_creacion), style_date)
                        if incidente.incidente.ordentrabajo.estado == 2:
                            fecha = str(incidente.incidente.ordentrabajo.fecha_modificacion)
                        else:
                            fecha = 'NO ESTA CERRADO'
                        ws.write(row_num, 4, "%s" % fecha, styrow)
                        ws.write(row_num, 5,
                                 incidente.incidente.ultimo_agente_asignado().persona.nombre_completo_inverso() if incidente.incidente.ultimo_agente_asignado() else '',
                                 styrow)

                        row_num += 1
                    row_num += 1
                    ws.write(row_num, 3, 'TOTAL', styrow)
                    ws.write(row_num, 4, incidentes.count(), styrow)
                    wb.save(response)
                    return response
                except Exception as ex:
                    pass

            elif action == 'reportelicencias':
                try:
                    __author__ = 'Unemi'
                    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
                    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                    title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    style1 = easyxf(num_format_str='D-MMM-YY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style.alignment.horz = font_style.alignment.HORZ_CENTER
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    font_style2.alignment.horz = font_style2.alignment.HORZ_CENTER
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('exp_xls_post_part')
                    ws.write_merge(0, 0, 0, 7, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    ws.write_merge(1, 1, 0, 2, 'De ' +request.GET['desde']+' hasta '+ request.GET['hasta'], style_sb)
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=Licencias_Instaladas ' + random.randint(1, 10000).__str__() + '.xls'
                    columns = [
                        (u"CODIGO INCIDENTE", 6000),
                        (u"ACTIVO TECNOLÓGICO", 20000),
                        (u"FECHA", 3000),
                        (u"SOLICITANTE", 10000),
                        (u"ESTADO DE INCIDENTE", 6000),
                        (u"ASUNTO", 10000),
                        (u"PRODUCTO", 8000),
                        (u"TIPO", 4000),
                        (u"ESTADO DE NOTIFICACIÓN", 7000),
                        (u"AGENTE", 10000),
                    ]
                    row_num = 3
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    date_format = xlwt.XFStyle()
                    date_format.num_format_str = 'yyyy/mm/dd'
                    incidenteprodet = HdIncidenteProductoDetalle.objects.filter(status=True, producto__grupo_id=26,  incidente__fechareporte__range=(request.GET['desde'], request.GET['hasta'] )).order_by('incidente__fechareporte')
                    row_num = 4
                    for prodet in incidenteprodet:
                        ws.write(row_num, 0, prodet.incidente.id, font_style2)
                        ws.write(row_num, 1, str(prodet.activo), font_style2)
                        ws.write(row_num, 2, prodet.incidente.fechareporte.strftime('%d-%m-%Y') , font_style2)
                        ws.write(row_num, 3, str(prodet.incidente.persona), font_style2)
                        ws.write(row_num, 4, str(prodet.incidente.estado), font_style2)
                        ws.write(row_num, 5, prodet.incidente.asunto, font_style2)
                        ws.write(row_num, 6, str(prodet.producto), font_style2)
                        ws.write(row_num, 7, str(prodet.producto.grupo) , font_style2)
                        ws.write(row_num, 8, prodet.get_estado_display() , font_style2)
                        ws.write(row_num, 9, str(prodet.incidente.ultimo_agente_asignado())  if prodet.incidente.ultimo_agente_asignado() else '', font_style2)
                        row_num += 1
                    wb.save(response)
                    return response
                except Exception as ex:
                    pass

            elif action == 'excellistaencuesta':
                try:
                    __author__ = 'Unemi'
                    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
                    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                    title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    style1 = easyxf(num_format_str='D-MMM-YY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('exp_xls_post_part')
                    ws.write_merge(0, 0, 0, 7, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    ws.write_merge(1, 1, 0, 2, 'De ' +request.GET['ini']+' hasta '+ request.GET['fin'], style_sb)
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=Resultadosencuesta ' + random.randint(1, 10000).__str__() + '.xls'
                    columns = [
                        (u"PREGUNTA", 15000),
                        (u"OBSERVACIÓN", 15000),
                        (u"RESPUESTA", 8000),
                    ]
                    row_num = 3
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    date_format = xlwt.XFStyle()
                    date_format.num_format_str = 'yyyy/mm/dd'
                    grupo = HdDetalle_Grupo.objects.get(estado=True, persona=persona).grupo
                    incidentes = HdRespuestaEncuestas.objects.filter(cabrespuesta__incidente__h_d_incidentes__agente__grupo=grupo,cabrespuesta__incidente__h_d_incidentes__estado_id=3,cabrespuesta__incidente__fechareporte__range=(convertir_fecha(request.GET['ini']), convertir_fecha(request.GET['fin'])), cabrespuesta__incidente__estado=3).distinct()
                    # incidentes = HdIncidente.objects.filter(status=True, fechareporte__range=(convertir_fecha(request.GET['de']), convertir_fecha(request.GET['hasta']))).order_by('fechareporte')
                    row_num = 4
                    for incidente in incidentes:
                        ws.write(row_num, 0, incidente.detencuesta.pregunta.nombre, font_style2)
                        ws.write(row_num, 1, incidente.observaciones, font_style2)
                        ws.write(row_num, 2, incidente.respuesta.nombre, font_style2)
                        row_num += 1
                    wb.save(response)
                    return response
                except Exception as ex:
                    pass
            # CAMBIOS DE JUSSY
            elif action == 'bodega':
                try:
                    data['title'] = u'Gestión de productos bodega'
                    search = None
                    ids = None
                    tipo = None
                    if 't' in request.GET and int(request.GET['t']) > 0:
                        data['tipoid'] = tipo = int(request.GET['t'])
                        productos = Producto.objects.filter(tipoproducto__id=tipo)
                    elif 's' in request.GET:
                        search = request.GET['s']
                        productos = Producto.objects.filter(Q(codigo__icontains=search) |
                                                            Q(descripcion__icontains=search) |
                                                            Q(cuenta__descripcion__icontains=search))
                    elif 'id' in request.GET:
                        ids = request.GET['id']
                        productos = Producto.objects.filter(id=ids)
                    else:
                        productos = Producto.objects.all()
                    paging = MiPaginador(productos, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['productos'] = page.object_list
                    data['tipos_productos'] = TipoProducto.objects.filter(status=True)
                    return render(request, "helpdesk_hdincidente/productos.html", data)
                except Exception as ex:
                    pass

            elif action == 'listaordentrabajo':
                try:
                    data['title'] = u'Orden de trabajo'
                    search = None
                    ids = None
                    tipo = None
                    ordenes = OrdenTrabajo.objects.filter(status=True)
                    if 's' in request.GET:
                        search = request.GET['s']
                        ordenes = ordenes.filter(Q(codigoorden__icontains=search))
                    elif 'id' in request.GET:
                        ids = request.GET['id']
                        ordenes = ordenes.filter(id=ids)
                    paging = MiPaginador(ordenes, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['ordenes'] = page.object_list
                    return render(request, "helpdesk_hdincidente/listaordentrabajo.html", data)
                except Exception as ex:
                    pass

            elif action == 'listamateriales':
                try:
                    data['title'] = u'Lista de Materiales'
                    search = None
                    ids = None
                    if 's' in request.GET:
                        search = request.GET['s']
                    if search:
                        materiales = HdMateriales.objects.filter(nombre__icontains=search, status=True)
                    elif 'id' in request.GET:
                        ids = request.GET['id']
                        materiales = HdMateriales.objects.filter(codigo=ids, status=True).order_by('nombre')
                    else:
                        materiales = HdMateriales.objects.filter(status=True).order_by('nombre')
                    paging = MiPaginador(materiales, 10)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['listamateriales'] = page.object_list
                    return render(request, "helpdesk_hdincidente/listamateriales.html", data)
                except Exception as ex:
                    pass

            elif action == 'cerrarorden':
                try:
                    data['title'] = u'Cerrar orden de trabajo'
                    data['orden'] = orden =  OrdenTrabajo.objects.get(pk=int(request.GET['id']))
                    data['detalleorden'] = orden.detalleordentrabajo_set.filter(status=True)
                    initial=  model_to_dict(orden)
                    data['form'] =  CerrarOrdenForm(initial=initial)
                    data['form2'] = DetalleOrdenForm()
                    return render(request, "helpdesk_hdincidente/cerrarorden.html", data)
                except Exception as ex:
                    pass
            elif action == 'ordenpdf':
                try:

                    #tecnico = Administrativo.objects.get(pk=int(request.GET['tecnico'])).persona.pk
                    desde = request.GET['de']
                    hasta = request.GET['hasta']
                    tec = int(request.GET['tecnico'])
                    bloque = int(request.GET['bloque'])
                    estado = int(request.GET['estado'])
                    checktodos = int(request.GET['checktodos'])
                    if checktodos == 1:
                        tec = 0
                    if not tec == 0:
                        tecnico = Administrativo.objects.get(pk=int(request.GET['tecnico'])).persona.pk
                        data['tecnico'] = Persona.objects.get(pk=int(tecnico)).nombre_completo_inverso()
                    else:
                        data['tecnico'] = 'TODOS LOS TECNICOS'
                    if not bloque == 0:
                        data['bloque'] = HdBloque.objects.get(pk=int(bloque)).nombre
                    else:
                        data['bloque'] = 'TODOS LOS BLOQUES'

                    # incidentes = HdDetalle_Incidente.objects.filter(incidente__ordentrabajo__fecha_creacion__range=(
                    # convertir_fecha(request.GET['de']), convertir_fecha(request.GET['hasta'])), incidente__status=True,
                    #                                                 agente__persona=tecnico,
                    #                                                 incidente__ubicacion__bloque=int(bloque),
                    #                                                 incidente__ordentrabajo__estado=int(estado))

                    if bloque == 0 and not estado == 0 and not tec == 0:
                        incidentes = HdDetalle_Incidente.objects.filter(incidente__ordentrabajo__fecha_creacion__range=(
                            convertir_fecha(request.GET['de']), convertir_fecha(request.GET['hasta'])),
                            agente__persona=int(tecnico), incidente__ordentrabajo__estado=int(estado),
                            incidente__status=True)
                    elif estado == 0 and not bloque == 0 and not tec == 0:
                        incidentes = HdDetalle_Incidente.objects.filter(incidente__ordentrabajo__fecha_creacion__range=(
                            convertir_fecha(request.GET['de']), convertir_fecha(request.GET['hasta'])),
                            incidente__status=True, agente__persona=int(tecnico),
                            incidente__ubicacion__bloque=int(bloque))
                    elif tec == 0 and not estado == 0 and not bloque == 0:
                        incidentes = HdDetalle_Incidente.objects.filter(incidente__ordentrabajo__fecha_creacion__range=(
                            convertir_fecha(request.GET['de']), convertir_fecha(request.GET['hasta'])),
                            incidente__status=True,
                            incidente__ubicacion__bloque=int(bloque), incidente__ordentrabajo__estado=int(estado))
                    elif tec == 0 and estado == 0 and bloque == 0:
                        incidentes = HdDetalle_Incidente.objects.filter(incidente__ordentrabajo__fecha_creacion__range=(
                            convertir_fecha(request.GET['de']), convertir_fecha(request.GET['hasta'])),
                            incidente__status=True)
                    elif estado == 0 and bloque == 0 and not tec == 0:
                        incidentes = HdDetalle_Incidente.objects.filter(incidente__ordentrabajo__fecha_creacion__range=(
                            convertir_fecha(request.GET['de']), convertir_fecha(request.GET['hasta'])),
                            incidente__status=True,
                            agente__persona=int(tecnico))
                    elif estado == 0 and tec == 0 and not bloque == 0:
                        incidentes = HdDetalle_Incidente.objects.filter(incidente__ordentrabajo__fecha_creacion__range=(
                            convertir_fecha(request.GET['de']), convertir_fecha(request.GET['hasta'])),
                            incidente__status=True,
                            incidente__ubicacion__bloque=int(bloque))
                    elif tec == 0 and bloque == 0 and not estado == 0:
                        incidentes = HdDetalle_Incidente.objects.filter(incidente__ordentrabajo__fecha_creacion__range=(
                            convertir_fecha(request.GET['de']), convertir_fecha(request.GET['hasta'])),
                            incidente__status=True,
                            incidente__ordentrabajo__estado=int(estado))
                    else:
                        incidentes = HdDetalle_Incidente.objects.filter(
                            incidente__ordentrabajo__fecha_creacion__range=(
                                convertir_fecha(request.GET['de']), convertir_fecha(request.GET['hasta'])),
                            incidente__status=True,
                            agente__persona=int(tecnico),
                            incidente__ubicacion__bloque=int(bloque),
                            incidente__ordentrabajo__estado=int(estado))

                    data['incidente'] = incidentes
                    data['total'] = incidentes.count()

                    esta = ''
                    if not estado == 0:
                        if estado == '1':
                            esta = 'GENERADO'
                        elif estado == '2':
                            esta = 'CERRADO'
                        elif estado == '3':
                            esta = 'PENDIENTE REPUESTO'
                        elif estado == '4':
                            esta = 'TALLER PARTICULAR'
                        elif estado == '5':
                            esta = 'EN TRÁMITE'
                        data['estado'] = esta
                    else:
                        data['estado'] = 'TODOS LOS ESTADOS'

                    #data['estado'] = esta
                    # data['bloque'] = HdBloque.objects.get(pk=int(bloque)).nombre
                    # data['tecnico'] = Persona.objects.get(pk=int(tecnico)).nombre_completo_inverso()
                    data['desde'] = desde
                    data['hasta'] = hasta

                    data['fechahoy'] = datetime.now().date()
                    return conviert_html_to_pdf(
                        'helpdesk_hdincidente/ordenpdf.html',
                        {
                            'pagesize': 'A4',
                            'data': data,
                        }
                    )
                except Exception as ex:
                    pass


            elif action == 'delorden':
                try:
                    data['title'] = u'Eliminar orden'
                    data['orden'] = OrdenTrabajo.objects.get(pk=int(request.GET['id']))
                    return render(request, "helpdesk_hdincidente/delorden.html", data)
                except Exception as ex:
                    pass

            return HttpResponseRedirect(request.path)
        else:
            try:
                data['title'] = u'Administrador Help Desk'
                tab = int(request.GET.get('t', 1))
                url_vars = f'&t={tab}'
                filtro = Q(status=True, tipoincidente_id=2)
                selecttipoeq, estadoOT, id_estado, idgrupo, search, ids = getNUM('idec'), getNUM('ideot'), getNUM('ide'), getNUM('idg'), getSTR('s'), getNUM('id')

                if search:
                    search = search.strip()
                    url_vars += f"&s={search}"
                    data['s'] = search
                    ss = search.split(' ')
                    if len(ss) == 1:
                        if selecttipoeq == 3:
                            filtro &= Q(ordentrabajo__codigoorden__icontains=search)
                        elif selecttipoeq == 0:
                            filtro &= Q(Q(h_d_incidentes__agente__persona__usuario__username__icontains=search) | Q(
                                ordentrabajo__codigoorden__icontains=search) | Q(
                                tipoincidente__nombre__icontains=search) | Q(asunto__icontains=search) | Q(
                                pk__icontains=search) | Q(persona__apellido1__icontains=search) | Q(persona__apellido2__icontains=search) | Q(persona__nombres__icontains=search))
                        else:
                            filtro &= Q(Q(h_d_incidentes__agente__persona__usuario__username__icontains=search) | Q(
                                tipoincidente__nombre__icontains=search) | Q(asunto__icontains=search) | Q(
                                pk__icontains=search) | Q(persona__apellido1__icontains=search) | Q(persona__apellido2__icontains=search) | Q(persona__nombres__icontains=search))

                    elif len(ss) == 2:
                        if selecttipoeq == 3:
                            filtro &= Q(ordentrabajo__codigoorden__icontains=search)
                        elif selecttipoeq == 0:
                            filtro &= Q(Q(h_d_incidentes__agente__persona__usuario__username__icontains=search) | Q(
                                ordentrabajo__codigoorden__icontains=search) | Q(
                                tipoincidente__nombre__icontains=search) | Q(asunto__icontains=search) | Q(
                                pk__icontains=search) | Q(persona__apellido1__icontains=search) | Q(persona__apellido2__icontains=search) | Q(persona__nombres__icontains=search) | Q(Q(persona__apellido1__icontains=ss[0]) & Q(persona__apellido2__icontains=ss[1])))
                        else:
                            filtro &= Q(Q(h_d_incidentes__agente__persona__usuario__username__icontains=search) | Q(
                                tipoincidente__nombre__icontains=search) | Q(asunto__icontains=search) | Q(
                                pk__icontains=search) | Q(persona__apellido1__icontains=search) | Q(persona__apellido2__icontains=search) | Q(persona__nombres__icontains=search) | Q(Q(persona__apellido1__icontains=ss[0]) & Q(persona__apellido2__icontains=ss[1])))

                    else:
                        if selecttipoeq == 3:
                            filtro &= Q(ordentrabajo__codigoorden__icontains=search)
                        elif selecttipoeq == 0:
                            filtro &= Q(Q(h_d_incidentes__agente__persona__usuario__username__icontains=search) | Q(
                                ordentrabajo__codigoorden__icontains=search) | Q(
                                tipoincidente__nombre__icontains=search) | Q(asunto__icontains=search) | Q(
                                pk__icontains=search) | Q(persona__apellido1__icontains=search) | Q(persona__apellido2__icontains=search) | Q(persona__nombres__icontains=search) | Q(Q(persona__apellido1__icontains=ss[0]) & Q(persona__apellido2__icontains=ss[1]) & Q(persona__nombres__icontains=ss[2])))
                        else:
                            filtro &= Q(Q(h_d_incidentes__agente__persona__usuario__username__icontains=search) | Q(
                                tipoincidente__nombre__icontains=search) | Q(asunto__icontains=search) | Q(
                                pk__icontains=search) | Q(persona__apellido1__icontains=search) | Q(persona__apellido2__icontains=search) | Q(persona__nombres__icontains=search) | Q(Q(persona__apellido1__icontains=ss[0]) & Q(persona__apellido2__icontains=ss[1]) & Q(persona__nombres__icontains=ss[2])))


                if selecttipoeq and not selecttipoeq > 4:
                    url_vars += f"&idec={selecttipoeq}"
                    data['selecttipoeq'] = selecttipoeq
                    filtro &= [None, Q(revisionequiposincodigo=True), Q(revisionequipoexterno=True), Q(ordentrabajo__codigoorden__isnull=False), Q(ordentrabajo__codigoorden__isnull=True)][selecttipoeq]

                if id_estado:
                    url_vars += f"&ide={id_estado}"
                    data['estadoid'] = id_estado
                    filtro &= Q(estado_id=id_estado)

                if ids:
                    data['ids'] = ids
                    filtro &= Q(id=ids)

                if estadoOT:
                    url_vars += f"&ideot={estadoOT}"
                    data['estadoOT'] = estadoOT
                    filtro &= Q(ordentrabajo__estado=estadoOT)

                if tab == 1:
                    paging = MiPaginador(HdIncidente.objects.filter(filtro).order_by('estado', '-id', 'fechareporte').distinct('estado', 'id'), 15)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['incidentescab'] = page.object_list
                    data['url_vars'] = url_vars


                # Mis incidentes
                data['puede_resolver_incidente'] = es_agente = persona.hdpersona.filter(status=True, estado=True)

                if es_agente and tab == 2:
                    url_vars += '&tab=2'
                    agente = HdDetalle_Grupo.objects.filter(persona=persona, status=True).first()
                    filtro &= Q(Q(status=True) & Q(h_d_incidentes__agente=agente) & Q(h_d_incidentes__estadoasignacion=1))
                    data['mis_incidentes_count'] = HdIncidente.objects.filter(Q(Q(estado_id=1) & Q(status=True) & Q(h_d_incidentes__agente=agente) & Q(h_d_incidentes__estadoasignacion=1))).values('id').count()

                    paging = MiPaginador(HdIncidente.objects.filter(filtro).order_by('estado', '-id', 'fechareporte').distinct('estado', 'id'), 15)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['misincidentes'] = page.object_list
                    data['url_vars'] = url_vars

                data['cargo'] = '(' + persona.distributivopersona_set.get(regimenlaboral_id=1, status=True, estadopuesto=1).denominacionpuesto.descripcion + ')' if persona.distributivopersona_set.filter(regimenlaboral_id=1, status=True, estadopuesto=1) else ''
                data['estados'] = HdEstado.objects.filter(status=True).exclude(pk=4)
                data['form2'] = SeleccionarActivoForm()
                data['reporte_0'] = obtener_reporte('reporte_incidente')
                data['bloque'] = HdBloque.objects.filter(status=True)
                data['estado'] = ESTADO_ORDEN_TRABAJO
                data['fecha'] = hoy
                data['grupos'] = HdTipoIncidente.objects.values_list('id', 'nombre', flat=False).filter(status=True, pk__in=[1, 2, 3])
                data['modules'] = panel()
                data['idgrupo'] = 2
                data['module_type'] = MODULE_TYPE
                data['cantincidente'] = HdIncidente.objects.filter(filtro).filter(estado_id=1).values('id').count()
                data['tab'] = tab
                return render(request, 'helpdesk_hdincidente/view.html', data)
            except Exception as ex:
                pass


MODULE_TYPE = ((0, ''), (1, 'Gestión de Usuario'), (2, 'Administración'))


def panel():
    return [
        {'name': u'Grupo sistemas/equipo', 'url': r'/helpdesk_hdplanificacion?action=viewgruposistemas', 'icon': u'bi bi-boxes', 'type': 0},
        {'name': u'Bienes', 'url': r'/helpdesk_hdplanificacion?action=viewbien', 'icon': u'bi bi-pc-display-horizontal', 'type': 0},
        {'name': u'Configuración frecuencia', 'url': r'/helpdesk_hdplanificacion?action=viewconf', 'icon': u'bi bi-calendar-plus', 'type': 0},
        {'name': u'Frecuencia', 'url': r'/helpdesk_hdplanificacion?action=viewfrecuencia', 'icon': u'bi bi-calendar-week', 'type': 0},
        {'name': u'Tipo de bien', 'url': r'/helpdesk_hdplanificacion?action=viewtipobien', 'icon': u'bi bi-pc-display', 'type': 0},
        {'name': u'Tareas de limpieza','url': r'/helpdesk_hdplanificacion?action=viewtlimpieza', 'icon': u'bi bi-brush', 'type': 0},
        {'name': u'Lista de daños', 'url': r'/helpdesk_hdplanificacion?action=viewdanios', 'icon': u'bi bi-bug', 'type': 0},
        {'name': u'Pieza/partes', 'url': r'/helpdesk_hdplanificacion?action=viewpiezaparte', 'icon': u'bi bi-diagram-3', 'type': 0},
        {'name': u'Reparación', 'url': r'/helpdesk_hdplanificacion?action=viewreparacion', 'icon': u'bi bi-bandaid', 'type': 0},
        {'name': u'Materiales mantenimiento', 'url': r'/helpdesk_hdplanificacion?action=viewmateriales', 'icon': u'bi bi-wrench-adjustable-circle', 'type': 0},
        {'name': u'Presupuesto', 'url': r'/helpdesk_hdplanificacion?action=viewpresupuesto', 'icon': u'bi bi-cash-coin', 'type': 0},
        {'name': u'Cronograma semanal', 'url': r'/helpdesk_hdplanificacion?action=viewgruposistemas', 'icon': u'bi bi-clock-history', 'type': 0},
        {'name': u'Grupos', 'url': r'/helpdesk_hdincidente?action=grupo', 'icon': u'fa fa-group', 'type': 1},
        {'name': u'Director', 'url': r'/helpdesk_hdincidente?action=viewdirector', 'icon': u'fa fa-group', 'type': 1},
        {'name': u'Urgencia', 'url': r'/helpdesk_hdincidente?action=urgencia', 'icon': u'fa fa-hourglass-start', 'type': 2},
        {'name': u'Impacto', 'url': r'/helpdesk_hdincidente?action=impacto', 'icon': u'fa fa-stack-overflow', 'type': 2},
        {'name': u'Prioridad', 'url': r'/helpdesk_hdincidente?action=prioridad', 'icon': u'fa fa-clock-o', 'type': 2},
        {'name': u'Tabla de Prioridad', 'url': r'/helpdesk_hdincidente?action=unionimpactourgenciaprioridad', 'icon': u'fa fa-object-group', 'type': 2},
        {'name': u'Estados', 'url': r'/helpdesk_hdincidente?action=estado', 'icon': u'fa fa-tags', 'type': 2},
        {'name': u'Categorias', 'url': r'/helpdesk_hdincidente?action=categoria', 'icon': u'fa fa-share-alt', 'type': 2},
        {'name': u'Medio de Reporte', 'url': r'/helpdesk_hdincidente?action=medioreporte', 'icon': u'fa fa-external-link', 'type': 2},
        {'name': u'Proceso de Incidente', 'url': r'/helpdesk_hdincidente?action=viewproceso', 'icon': u'fa fa-spinner', 'type': 2},
        {'name': u'Bloques', 'url': r'/helpdesk_hdincidente?action=bloque_ubicacion', 'icon': u'fa fa-cubes', 'type': 2},
        {'name': u'Causas de Incidentes', 'url': r'/helpdesk_hdincidente?action=listadocausas', 'icon': u'fa fa-list', 'type': 2},
        {'name': u'Piezas y Partes', 'url': r'/helpdesk_hdincidente?action=piezapartes', 'icon': u'fa fa-cogs', 'type': 2},
        {'name': u'Solicitudes Piezas y Partes', 'url': r'/helpdesk_hdincidente?action=solicitudespiezapartes', 'icon': u'fa fa-cogs', 'type': 2},
        {'name': u'Tipo de incidente', 'url': r'/helpdesk_hdincidente?action=tipoincidente', 'icon': u'fa fa-sitemap', 'type': 2},
        {'name': u'Existencia Bodega', 'url': r'/helpdesk_hdincidente?action=bodega', 'icon': u'fa fa-archive', 'type': 2},
        {'name': u'Orden Trabajo', 'url': r'/helpdesk_hdincidente?action=listaordentrabajo', 'icon': u'fa fa-ambulance', 'type': 2},
        {'name': u'Materiales', 'url': r'/helpdesk_hdincidente?action=listamateriales', 'icon': u'fa fa-list', 'type': 2}
    ]


def _notificar_reasignacion_incidente(**kwargs):
    from webpush.utils import _send_notification
    from wpush.models import SubscriptionInfomation
    try:
        persona = kwargs.pop('persona')
        incidente = kwargs.pop('incidente')
        subscriptions = persona.usuario.webpush_info.select_related("subscription")
        push_infos = SubscriptionInfomation.objects.filter(subscription_id__in=subscriptions.values_list('subscription__id', flat=True), app__in=[1,2], status=True).select_related("subscription")
        url = f"/helpdesk_hdincidente?id={incidente.pk}"
        titulo = "Reasignación de incidente TIC"
        cuerpo = f""""Hola {persona.nombres.split(' ')[0]}, \nTe informamos que se te ha asignado el incidente tecnológico {incidente} y requiere tu atención lo antes posible."""

        if not push_infos:
            notificacion2(titulo, cuerpo, persona, None, url, incidente.pk, 1, 'sga', incidente)
            return 0

        for device in push_infos:
            payload = {
                "head": titulo,
                "body": cuerpo,
                "url": url,
                "action": "loadNotifications",
            }
            try:
                _send_notification(device.subscription, json.dumps(payload), ttl=500)
            except Exception as ex:
                notificacion2(titulo, cuerpo, persona, None, url, incidente.pk, 1, 'sga', incidente)
                return 0

        return 1
    except Exception as ex:
        print(f"Fallo de envio del push notification: {ex.__str__()}")
        return 0