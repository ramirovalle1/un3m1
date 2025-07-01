# coding=utf-8
import json
import os
import io as StringIO

from django.core.files.base import ContentFile
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.lib.colors import HexColor
from reportlab.lib.units import inch

from settings import MEDIA_ROOT, MEDIA_URL, SITE_STORAGE, SITE_ROOT, DEBUG
from django.template.loader import get_template
from django.template import Context
from xhtml2pdf import pisa
import settings
from django.http import HttpResponse, JsonResponse

# ----------------------------------REPORLAB----------------------------------------------------------------

from reportlab.platypus import Paragraph, SimpleDocTemplate, TableStyle, Image
from reportlab.lib.styles import StyleSheet1, ParagraphStyle, LineStyle
from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER, TA_JUSTIFY
from reportlab.lib import colors
from xhtml2pdf.default import DEFAULT_FONT, DEFAULT_CSS
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont, TTFOpenFile
from reportlab.graphics.shapes import Drawing
from io import BytesIO
from PyPDF2 import PdfFileMerger
from sga.funciones import null_to_numeric


def link_callback(uri, rel):
    sUrl = '/static/'      # Typically /static/
    sRoot = '/var/lib/django/academico/static/' if not DEBUG else os.path.join(SITE_ROOT, 'static')    # Typically /home/userX/project_static/
    mUrl = settings.MEDIA_URL       # Typically /static/media/
    mRoot = os.path.join(SITE_STORAGE, 'media') if not DEBUG else os.path.join(SITE_ROOT, 'media')     # Typically /home/userX/project_static/media/

    if uri.startswith(mUrl):
        path = os.path.join(mRoot, uri.replace(mUrl, ""))
    elif uri.startswith(sUrl):
        path = os.path.join(sRoot, uri.replace(sUrl, ""))
    else:
        return uri                  # handle absolute uri (ie: http://some.tld/foo.png)
    print(path)
    # make sure that file exists
    if not os.path.isfile(path):
        raise Exception('media URI must start with %s or %s' % (sUrl, mUrl))

    return path


def conviert_html_to_pdf(template_src, context_dict):
    template = get_template(template_src)
    html = template.render(context_dict).encode(encoding="UTF-8")
    result = StringIO.BytesIO()
    pisaStatus = pisa.CreatePDF(StringIO.BytesIO(html), result, link_callback=link_callback)
    if not pisaStatus.err:
        print(pisaStatus.err)
        return HttpResponse(result.getvalue(), content_type='application/pdf')
    return JsonResponse({"result": "bad", "mensaje": u"Problemas al ejecutar el reporte. %s" % pisaStatus.err})


def conviert_html_to_2pdf(template_src, context_dict, file_name):
    template = get_template(template_src)
    html = template.render(context_dict).encode(encoding="UTF-8")
    result = BytesIO()
    pisaStatus = pisa.CreatePDF(BytesIO(html), result)
    if not pisaStatus.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{file_name}.pdf"'
        return response
    return JsonResponse({"result": "bad", "mensaje": f"Problemas al ejecutar el reporte. {pisaStatus.err}"})


def download_html_to_pdf(template_src, context_dict):
    template = get_template(template_src)
    html = template.render(context_dict).encode(encoding="UTF-8")
    result = StringIO.BytesIO()
    pisaStatus = pisa.CreatePDF(StringIO.BytesIO(html), result, link_callback=link_callback)
    if not pisaStatus.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = "attachment; filename=archivo.pdf"
        return response
    return JsonResponse({"result": "bad", "mensaje": u"Problemas al ejecutar el reporte. %s" % pisaStatus.err})

def download_html_to_pdf_get_content(template_src, context_dict):
    template = get_template(template_src)
    html = template.render(context_dict).encode(encoding="UTF-8")
    result = StringIO.BytesIO()
    pisaStatus = pisa.CreatePDF(StringIO.BytesIO(html), result, link_callback=link_callback)
    if not pisaStatus.err:
        return result.getvalue()
    return JsonResponse({"result": "bad", "mensaje": u"Problemas al ejecutar el reporte. %s" % pisaStatus.err})


def conviert_html_to_pdf_save_file_model(template_src, context_dict, name=None):
    template = get_template(template_src)
    html = template.render(context_dict).encode(encoding="UTF-8")
    result = StringIO.BytesIO()
    pisa.CreatePDF(html, dest=result, encoding='utf-8')
    pdf_file = ContentFile(result.getvalue(), name=name)
    # Muestra el PDF en el navegador
    response = HttpResponse(result.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="documento.pdf"'
    return pdf_file, response


def conviert_html_to_pdf_name(template_src, context_dict, namefile):
    template = get_template(template_src)
    html = template.render(context_dict).encode(encoding="UTF-8")
    result = StringIO.BytesIO()
    pisaStatus = pisa.CreatePDF(StringIO.BytesIO(html), result, link_callback=link_callback)
    if not pisaStatus.err:
        print(pisaStatus.err)
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename={}.pdf'.format(namefile)
        return response
    return JsonResponse({"result": "bad", "mensaje": u"Problemas al ejecutar el reporte. %s" % pisaStatus.err})


def conviert_html_to_pdf_name_bitacora(template_src, context_dict, namefile):
    template = get_template(template_src)
    html = template.render(context_dict).encode(encoding="UTF-8")
    result = StringIO.BytesIO()
    pisaStatus = pisa.CreatePDF(StringIO.BytesIO(html), result, link_callback=link_callback)
    if not pisaStatus.err:
        return True,result
    return False, JsonResponse({"result": True, "mensaje": u"Problemas al ejecutar el reporte. %s" % pisaStatus.err})


def conviert_html_to_pdf_name_save(template_src, context_dict, namefile):
    template = get_template(template_src)
    html = template.render(context_dict).encode(encoding="UTF-8")
    result = StringIO.BytesIO()
    output_folder = os.path.join(SITE_STORAGE, 'media', 'qrcode', 'solicitudempresas')
    filepdf = open(output_folder + os.sep + "{}.pdf".format(namefile), "w+b")
    links = lambda uri, rel: os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, ''))
    pdf1 = pisa.pisaDocument(StringIO.BytesIO(html), dest=filepdf, link_callback=links)
    pisaStatus = pisa.CreatePDF(StringIO.BytesIO(html), result, link_callback=links)
    if not pdf1.err:
        # return HttpResponse(result.getvalue(), content_type='application/pdf')
        return True
    return JsonResponse({"result": "bad", "mensaje": u"Problemas al ejecutar el reporte."})


def conviert_html_to_pdf_name_savecartavinc(template_src, context_dict, namefile):
    template = get_template(template_src)
    html = template.render(context_dict).encode(encoding="UTF-8")
    result = StringIO.BytesIO()
    output_folder = os.path.join(SITE_STORAGE, 'media', 'qrcode', 'solicitudempresascarta')
    filepdf = open(output_folder + os.sep + "{}.pdf".format(namefile), "w+b")
    links = lambda uri, rel: os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, ''))
    pdf1 = pisa.pisaDocument(StringIO.BytesIO(html), dest=filepdf, link_callback=links)
    pisaStatus = pisa.CreatePDF(StringIO.BytesIO(html), result, link_callback=links)
    if not pdf1.err:
        # return HttpResponse(result.getvalue(), content_type='application/pdf')
        return True
    return JsonResponse({"result": "bad", "mensaje": u"Problemas al ejecutar el reporte."})

def conviert_html_to_pdf_name_saveaccionpersonal(template_src, context_dict, namefile):
    template = get_template(template_src)
    html = template.render(context_dict).encode(encoding="UTF-8")
    result = StringIO.BytesIO()
    output_folder = os.path.join(SITE_STORAGE, 'media', 'accionpersonal')
    filepdf = open(output_folder + os.sep + "{}.pdf".format(namefile), "w+b")
    links = lambda uri, rel: os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, ''))
    pdf1 = pisa.pisaDocument(StringIO.BytesIO(html), dest=filepdf, link_callback=links)
    pisaStatus = pisa.CreatePDF(StringIO.BytesIO(html), result, link_callback=links)
    if not pdf1.err:
        # return HttpResponse(result.getvalue(), content_type='application/pdf')
        return True
    return JsonResponse({"result": "bad", "mensaje": u"Problemas al ejecutar el reporte."})


def convert_html_to_pdf(template_src, context_dict, filename, directoryname):
    template = get_template(template_src)
    html = template.render(context_dict).encode(encoding="UTF-8")
    result = StringIO.BytesIO()
    output_folder = os.path.join(SITE_STORAGE, directoryname)
    filepdf = open(output_folder + os.sep + filename, "w+b")
    links = lambda uri, rel: os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, ''))
    pdf1 = pisa.pisaDocument(StringIO.BytesIO(html), dest=filepdf, link_callback=links)
    pisaStatus = pisa.CreatePDF(StringIO.BytesIO(html), result, link_callback=links)
    if not pdf1.err:
        # return HttpResponse(result.getvalue(), content_type='application/pdf')
        return True
    else:
        return False
    # return JsonResponse({"result": "bad", "mensaje": u"Problemas al ejecutar el reporte."})


def convert_html_to_excel(template_src, context_dict, filename, directoryname):
    import os
    import pandas as pd
    from bs4 import BeautifulSoup
    from openpyxl import Workbook as openxl
    from openpyxl.styles import Font as openxlFont
    from openpyxl.styles.alignment import Alignment as alin
    from django.template.loader import get_template
    from openpyxl.utils.dataframe import dataframe_to_rows
    try:
        # Renderizar la plantilla HTML con el contexto proporcionado
        template = get_template(template_src)
        html = template.render(context_dict)

        # Parsear el HTML con BeautifulSoup
        soup = BeautifulSoup(html, 'html.parser')

        # Encontrar la tabla en el HTML
        table = soup.find('table')

        # Convertir la tabla HTML a un DataFrame de pandas
        df = pd.read_html(str(table))[0]

        # Crear un nuevo libro de trabajo y una hoja
        wb = openxl()
        ws = wb.active
        ws.title = "Resultados"

        # Escribir los encabezados
        for col_num, column_title in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_num, value=column_title)
            cell.font = openxlFont(bold=True)
            cell.alignment = alin(horizontal="center", vertical="center")

        # Escribir los datos
        for row in dataframe_to_rows(df, index=False, header=False):
            ws.append(row)

            # Eliminar la última fila
        ws.delete_rows(ws.max_row)

        # Formatear los porcentajes correctamente
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if isinstance(cell.value, str) and '%' in cell.value:
                    cell.value = float(cell.value.replace('%', '').replace(',', '.')) / 100
                    cell.number_format = '0.00%'

        # Ajustar automáticamente el ancho de las columnas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Obtiene la letra de la columna
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # Guardar el archivo
        # Crear el directorio de salida si no existe
        output_folder = os.path.join(SITE_STORAGE, directoryname)
        os.makedirs(output_folder, exist_ok=True)

        # Guardar el archivo Excel
        output_path = os.path.join(output_folder, filename)
        wb.save(output_path)

        return True
    except Exception as e:
        pass

def conviert_html_to_pdfsave(template_src, context_dict, filename):
    template = get_template(template_src)
    html = template.render(context_dict).encode(encoding="UTF-8")
    result = StringIO.BytesIO()
    output_folder = os.path.join(SITE_STORAGE, 'media', 'qrcode', 'evaluaciondocente')
    filepdf = open(output_folder + os.sep + filename, "w+b")
    links = lambda uri, rel: os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, ''))
    pdf1 = pisa.pisaDocument(StringIO.BytesIO(html), dest=filepdf, link_callback=links)
    pisaStatus = pisa.CreatePDF(StringIO.BytesIO(html), result, link_callback=links)
    if not pdf1.err:
        return HttpResponse(result.getvalue(), content_type='application/pdf')
    return JsonResponse({"result": "bad", "mensaje": u"Problemas al ejecutar el reporte."})

def conviert_html_to_pdfsavevistaprevia(template_src, context_dict, filename):
    template = get_template(template_src)
    html = template.render(context_dict).encode(encoding="UTF-8")
    result = StringIO.BytesIO()
    output_folder = os.path.join(SITE_STORAGE, 'media', 'qrcode', 'certificados')
    filepdf = open(output_folder + os.sep + filename, "w+b")
    links = lambda uri, rel: os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, ''))
    pdf1 = pisa.pisaDocument(StringIO.BytesIO(html), dest=filepdf, link_callback=links)
    pisaStatus = pisa.CreatePDF(StringIO.BytesIO(html), result, link_callback=links)
    if not pdf1.err:
        return HttpResponse(result.getvalue(), content_type='application/pdf')
    return JsonResponse({"result": "bad", "mensaje": u"Problemas al ejecutar el reporte."})

def conviert_html_to_pdfsavevistaprevia_2(template_src, context_dict, filename):
    template = get_template(template_src)
    html = template.render(context_dict).encode(encoding="UTF-8")
    result = StringIO.BytesIO()
    output_folder = os.path.join(SITE_STORAGE, 'media', 'qrcode', 'certificados_fe')
    filepdf = open(output_folder + os.sep + filename, "w+b")
    links = lambda uri, rel: os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, ''))
    pdf1 = pisa.pisaDocument(StringIO.BytesIO(html), dest=filepdf, link_callback=links)
    pisaStatus = pisa.CreatePDF(StringIO.BytesIO(html), result, link_callback=links)
    if not pdf1.err:
        return HttpResponse(result.getvalue(), content_type='application/pdf')
    return JsonResponse({"result": "bad", "mensaje": u"Problemas al ejecutar el reporte."})


# NO VALIDA 
# def conviert_html_to_pdf(template_src, context_dict, filename):
#     template = get_template(template_src)
#     html = template.render(context_dict).encode(encoding="UTF-8")
#     result = StringIO.BytesIO()
#     output_folder = os.path.join(SITE_STORAGE, 'media', 'qrcode', 'evaluaciondocente')
#     filepdf = open(output_folder + os.sep + filename, "w+b")
#     links = lambda uri, rel: os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, ''))
#     pdf1 = pisa.pisaDocument(StringIO.BytesIO(html), dest=filepdf, link_callback=links)
#     pisaStatus = pisa.CreatePDF(StringIO.BytesIO(html), result, link_callback=links)
#     if not pdf1.err:
#         return HttpResponse(result.getvalue(), content_type='application/pdf')
#     return JsonResponse({"result": "bad", "mensaje": u"Problemas al ejecutar el reporte."})

def conviert_html_to_pdfsavecertificados(template_src, context_dict, filename, rutacarpeta):
    template = get_template(template_src)
    html = template.render(context_dict).encode(encoding="UTF-8")
    result = StringIO.BytesIO()
    output_folder = os.path.join(SITE_STORAGE, 'media', 'qrcode', rutacarpeta)
    filepdf = open(output_folder + os.sep + filename, "w+b")
    links = lambda uri, rel: os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, ''))
    pdf1 = pisa.pisaDocument(StringIO.BytesIO(html), dest=filepdf, link_callback=links)
    pisaStatus = pisa.CreatePDF(StringIO.BytesIO(html), result, link_callback=links)
    if not pdf1.err:
        return HttpResponse(result.getvalue(), content_type='application/pdf')
    return JsonResponse({"result": "bad", "mensaje": u"Problemas al ejecutar el reporte."})

def conviert_html_to_pdfsaveqrcertificado(template_src, context_dict, filename):
    template = get_template(template_src)
    html = template.render(context_dict).encode(encoding="UTF-8")
    result = StringIO.BytesIO()
    output_folder = os.path.join(SITE_STORAGE, 'media', 'qrcode', 'certificados')
    filepdf = open(output_folder + os.sep + filename, "w+b")
    links = lambda uri, rel: os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, ''))
    pdf1 = pisa.pisaDocument(StringIO.BytesIO(html), dest=filepdf, link_callback=links)
    pisaStatus = pisa.CreatePDF(StringIO.BytesIO(html), result, link_callback=links)
    if not pdf1.err:
        # return HttpResponse(result.getvalue(), content_type='application/pdf')
        return True
    return JsonResponse({"result": "bad", "mensaje": u"Problemas al ejecutar el reporte."})

def conviert_html_to_pdfsaveqrtitulo(template_src, context_dict, filename, vistaprevia=False):
    template = get_template(template_src)
    html = template.render(context_dict).encode(encoding="UTF-8")
    result = StringIO.BytesIO()
    output_folder = os.path.join(SITE_STORAGE, 'media', 'qrcode', 'titulos')
    filepdf = open(output_folder + os.sep + filename, "w+b")
    links = lambda uri, rel: os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, ''))
    pdf1 = pisa.pisaDocument(StringIO.BytesIO(html), dest=filepdf, link_callback=links)
    pisaStatus = pisa.CreatePDF(StringIO.BytesIO(html), result, link_callback=links)
    if not pdf1.err:
        if vistaprevia:
            return HttpResponse(result.getvalue(), content_type='application/pdf')
        else:
            return True
    return JsonResponse({"result": "bad", "mensaje": u"Problemas al ejecutar el reporte."})

def conviert_html_to_pdfsaveqr_omacertificado(template_src, context_dict, filename):
    template = get_template(template_src)
    html = template.render(context_dict).encode(encoding="UTF-8")
    result = StringIO.BytesIO()
    output_folder = os.path.join(SITE_STORAGE, 'media', 'qrcode', 'omaCertificados')
    filepdf = open(output_folder + os.sep + filename, "w+b")
    links = lambda uri, rel: os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, ''))
    pdf1 = pisa.pisaDocument(StringIO.BytesIO(html), dest=filepdf, link_callback=links)
    pisaStatus = pisa.CreatePDF(StringIO.BytesIO(html), result, link_callback=links)
    if not pdf1.err:
            return True
    return JsonResponse({"result": "bad", "mensaje": u"Problemas al ejecutar el certificado."})

def conviert_html_to_pdfsaveinformeactivo(template_src, context_dict, filename):
    template = get_template(template_src)
    html = template.render(context_dict).encode(encoding="UTF-8")
    result = StringIO.BytesIO()
    output_folder = os.path.join(SITE_STORAGE, 'media', 'entregaactivo')
    filepdf = open(output_folder + os.sep + filename, "w+b")
    links = lambda uri, rel: os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, ''))
    pdf1 = pisa.pisaDocument(StringIO.BytesIO(html), dest=filepdf, link_callback=links)
    pisaStatus = pisa.CreatePDF(StringIO.BytesIO(html), result, link_callback=links)
    if not pdf1.err:
        # return HttpResponse(result.getvalue(), content_type='application/pdf')
        return True
    return JsonResponse({"result": "bad", "mensaje": u"Problemas al ejecutar el informe."})

def conviert_html_to_pdfsaveinformeinventarioactivostecnologicos(template_src, context_dict, filename):
    template = get_template(template_src)
    html = template.render(context_dict).encode(encoding="UTF-8")
    result = StringIO.BytesIO()
    output_folder = os.path.join(SITE_STORAGE, 'media', 'inventarioactivostecnologicos')
    filepdf = open(output_folder + os.sep + filename, "w+b")
    links = lambda uri, rel: os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, ''))
    pdf1 = pisa.pisaDocument(StringIO.BytesIO(html), dest=filepdf, link_callback=links)
    pisaStatus = pisa.CreatePDF(StringIO.BytesIO(html), result, link_callback=links)
    if not pdf1.err:
        # return HttpResponse(result.getvalue(), content_type='application/pdf')
        return True
    return JsonResponse({"result": "bad", "mensaje": u"Problemas al ejecutar el informe."})

def conviert_html_to_pdfsaveqrcertificado_v2(template_src, context_dict, filename):
    template = get_template(template_src)
    html = template.render(context_dict).encode(encoding="UTF-8")
    result = StringIO.BytesIO()
    output_folder = os.path.join(SITE_STORAGE, 'media', 'qrcode', 'certificados')
    filepdf = open(output_folder + os.sep + filename, "w+b")
    links = lambda uri, rel: os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, ''))
    pdf1 = pisa.pisaDocument(StringIO.BytesIO(html), dest=filepdf, link_callback=links)
    pisaStatus = pisa.CreatePDF(StringIO.BytesIO(html), result, link_callback=links)
    if not pdf1.err:
        # return HttpResponse(result.getvalue(), content_type='application/pdf')
        return True, pdf1, result
    return False, pdf1, None

def conviert_html_to_pdfsaveqrcertificadoscongresoinscrito(template_src, context_dict, filename):
    template = get_template(template_src)
    html = template.render(context_dict).encode(encoding="UTF-8")
    result = StringIO.BytesIO()
    output_folder = os.path.join(SITE_STORAGE, 'media', 'certificadoscongresoinscrito')
    filepdf = open(output_folder + os.sep + filename, "w+b")
    links = lambda uri, rel: os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, ''))
    pdfmetrics.registerFont(TTFont('zhfont', os.path.join(SITE_ROOT, 'static/fonts/Great_Vibes/GreatVibes-Regular.ttf')))
    DEFAULT_FONT["helvetica"] = "zhfont"
    # css = open(os.path.join(app.root_path, "static/css/pdf.css")).read()
    pdf1 = pisa.pisaDocument(StringIO.BytesIO(html), dest=filepdf, link_callback=links)
    pisaStatus = pisa.CreatePDF(StringIO.BytesIO(html), result, link_callback=links)
    if not pdf1.err:
        return HttpResponse(result.getvalue(), content_type='application/pdf')
        # return True
    return False

def conviert_html_to_pdfsaveqrcertificadoscongresoinscritoturistica(template_src, context_dict, output_folder, filename):
    template = get_template(template_src)
    html = template.render(context_dict).encode(encoding="UTF-8")
    result = StringIO.BytesIO()
    # output_folder = os.path.join(SITE_STORAGE, 'media', 'certificadoscongresoinscrito')
    filepdf = open(output_folder + os.sep + filename, "w+b")
    links = lambda uri, rel: os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, ''))
    pdfmetrics.registerFont(TTFont('zhfont', os.path.join(SITE_ROOT, 'static/fonts/Great_Vibes/GreatVibes-Regular.ttf')))
    DEFAULT_FONT["helvetica"] = "zhfont"
    # css = open(os.path.join(app.root_path, "static/css/pdf.css")).read()
    pdf1 = pisa.pisaDocument(StringIO.BytesIO(html), dest=filepdf, link_callback=links)
    pisaStatus = pisa.CreatePDF(StringIO.BytesIO(html), result, link_callback=links)
    if not pdf1.err:
        return {"isSuccess": True, "message": "", "data": {"filepdf": filepdf, "result": result}}
    return {"isSuccess": False, "message": "Ocurrio un error al generar pdf", "data": {"pdf1": pdf1}}


def conviert_html_to_pdfsavesilabo(template_src, context_dict, filename):
    template = get_template(template_src)
    html = template.render(context_dict).encode(encoding="UTF-8")
    result = StringIO.BytesIO()
    output_folder = os.path.join(SITE_STORAGE, 'media', 'documentos', 'silabos')
    filepdf = open(output_folder + os.sep + filename, "w+b")
    links = lambda uri, rel: os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, ''))
    pdf1 = pisa.pisaDocument(StringIO.BytesIO(html), dest=filepdf, link_callback=links)
    pisaStatus = pisa.CreatePDF(StringIO.BytesIO(html), result, link_callback=links)
    if not pdf1.err:
        # return HttpResponse(result.getvalue(), content_type='application/pdf')
        return True
    return JsonResponse({"result": "bad", "mensaje": u"Problemas al ejecutar el reporte."})

def conviert_html_to_pdfsavepracticas(template_src, context_dict, filename):
    template = get_template(template_src)
    html = template.render(context_dict).encode(encoding="UTF-8")
    result = StringIO.BytesIO()
    output_folder = os.path.join(SITE_STORAGE, 'media', 'documentos', 'guia_practicas')
    filepdf = open(output_folder + os.sep + filename, "w+b")
    links = lambda uri, rel: os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, ''))
    pdf1 = pisa.pisaDocument(StringIO.BytesIO(html), dest=filepdf, link_callback=links)
    pisaStatus = pisa.CreatePDF(StringIO.BytesIO(html), result, link_callback=links)
    if not pdf1.err:
        # return HttpResponse(result.getvalue(), content_type='application/pdf')
        return True
    return JsonResponse({"result": "bad", "mensaje": u"Problemas al ejecutar el reporte."})

def conviert_html_to_pdfsaveqrinformepracticasmensual(template_src, context_dict, filename):
    template = get_template(template_src)
    html = template.render(context_dict).encode(encoding="UTF-8")
    result = StringIO.BytesIO()
    output_folder = os.path.join(SITE_STORAGE, 'media', 'qrcode', 'informesppp')
    filepdf = open(output_folder + os.sep + filename, "w+b")
    links = lambda uri, rel: os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, ''))
    pdf1 = pisa.pisaDocument(StringIO.BytesIO(html), dest=filepdf, link_callback=links)
    pisaStatus = pisa.CreatePDF(StringIO.BytesIO(html), result, link_callback=links)
    if not pdf1.err:
        # return HttpResponse(result.getvalue(), content_type='application/pdf')
        return True
    return JsonResponse({"result": "bad", "mensaje": u"Problemas al ejecutar el reporte."})

def conviert_html_to_pdfsaveqrinformepractividadextra(template_src, context_dict, filename):
    template = get_template(template_src)
    html = template.render(context_dict).encode(encoding="UTF-8")
    result = StringIO.BytesIO()
    output_folder = os.path.join(SITE_STORAGE, 'media','convalidacionppv','informesactv')
    filepdf = open(output_folder + os.sep + filename, "w+b")
    links = lambda uri, rel: os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, ''))
    pdf1 = pisa.pisaDocument(StringIO.BytesIO(html), dest=filepdf, link_callback=links)
    pisaStatus = pisa.CreatePDF(StringIO.BytesIO(html), result, link_callback=links)
    if not pdf1.err:
        # return HttpResponse(result.getvalue(), content_type='application/pdf')
        return True
    return JsonResponse({"result": "bad", "mensaje": u"Problemas al ejecutar el reporte."})

def conviert_html_to_pdfsaveqrcertificadoinstructor(template_src, context_dict, filename):
    template = get_template(template_src)
    html = template.render(context_dict).encode(encoding="UTF-8")
    result = StringIO.BytesIO()
    output_folder = os.path.join(SITE_STORAGE, 'media', 'qrcode', 'certificados_facilitadores')
    filepdf = open(output_folder + os.sep + filename, "w+b")
    links = lambda uri, rel: os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, ''))
    pdf1 = pisa.pisaDocument(StringIO.BytesIO(html), dest=filepdf, link_callback=links)
    pisaStatus = pisa.CreatePDF(StringIO.BytesIO(html), result, link_callback=links)
    if not pdf1.err:
        return HttpResponse(result.getvalue(), content_type='application/pdf')
        #return True
    return JsonResponse({"result": "bad", "mensaje": u"Problemas al ejecutar el reporte."})

def conviert_html_to_pdfsaveqrcertificadocapacitacioninstructor(template_src, context_dict, filename):
    template = get_template(template_src)
    html = template.render(context_dict).encode(encoding="UTF-8")
    result = StringIO.BytesIO()
    output_folder = os.path.join(SITE_STORAGE, 'media', 'qrcode', 'certificados_capacitacion_facilitadores')
    filepdf = open(output_folder + os.sep + filename, "w+b")
    links = lambda uri, rel: os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, ''))
    pdf1 = pisa.pisaDocument(StringIO.BytesIO(html), dest=filepdf, link_callback=links)
    pisaStatus = pisa.CreatePDF(StringIO.BytesIO(html), result, link_callback=links)
    if not pdf1.err:
        return HttpResponse(result.getvalue(), content_type='application/pdf')
        #return True
    return JsonResponse({"result": "bad", "mensaje": u"Problemas al ejecutar el reporte."})

def conviert_html_to_pdfsaveqrcertificadoinstructor2(template_src, context_dict, filename):
    template = get_template(template_src)
    html = template.render(context_dict).encode(encoding="UTF-8")
    result = StringIO.BytesIO()
    output_folder = os.path.join(SITE_STORAGE, 'media', 'qrcode', 'certificados_facilitadores')
    filepdf = open(output_folder + os.sep + filename, "w+b")
    links = lambda uri, rel: os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, ''))
    pdf1 = pisa.pisaDocument(StringIO.BytesIO(html), dest=filepdf, link_callback=links)
    pisaStatus = pisa.CreatePDF(StringIO.BytesIO(html), result, link_callback=links)
    if not pdf1.err:
        #return HttpResponse(result.getvalue(), content_type='application/pdf')
        return True
    return JsonResponse({"result": "bad", "mensaje": u"Problemas al ejecutar el reporte."})

def conviert_html_to_pdfsavecartavinculacion(template_src, context_dict, filename):
    template = get_template(template_src)
    html = template.render(context_dict).encode(encoding="UTF-8")
    result = StringIO.BytesIO()
    output_folder = os.path.join(SITE_STORAGE, 'media', 'cartavinculacionpracticaspreprofesionales')
    filepdf = open(output_folder + os.sep + filename, "w+b")
    links = lambda uri, rel: os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, ''))
    pdf1 = pisa.pisaDocument(StringIO.BytesIO(html), dest=filepdf, link_callback=links)
    pisaStatus = pisa.CreatePDF(StringIO.BytesIO(html), result, link_callback=links)
    if not pdf1.err:
        return HttpResponse(result.getvalue(), content_type='application/pdf')
    return JsonResponse({"result": "bad", "mensaje": u"Problemas al ejecutar el reporte."})

def conviert_html_to_pdfsaveqrsilabo(template_src, context_dict, filename):
    template = get_template(template_src)
    html = template.render(context_dict).encode(encoding="UTF-8")
    result = StringIO.BytesIO()
    output_folder = os.path.join(SITE_STORAGE, 'media', 'qrcode', 'silabodocente')
    filepdf = open(output_folder + os.sep + filename, "w+b")
    links = lambda uri, rel: os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, ''))
    pdf1 = pisa.pisaDocument(StringIO.BytesIO(html), dest=filepdf, link_callback=links)
    pisaStatus = pisa.CreatePDF(StringIO.BytesIO(html), result, link_callback=links)
    if not pdf1.err:
        return True
    return JsonResponse({"result": "bad", "mensaje": u"Problemas al ejecutar el reporte."})

def conviert_html_to_pdfsaveadmitidos(template_src, context_dict, filename):
    template = get_template(template_src)
    html = template.render(context_dict).encode(encoding="UTF-8")
    result = StringIO.BytesIO()
    output_folder = os.path.join(SITE_STORAGE, 'media', 'qrcode', 'admitidos')
    filepdf = open(output_folder + os.sep + filename, "w+b")
    links = lambda uri, rel: os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, ''))
    pdf1 = pisa.pisaDocument(StringIO.BytesIO(html), dest=filepdf, link_callback=links)
    pisaStatus = pisa.CreatePDF(StringIO.BytesIO(html), result, link_callback=links)
    if not pdf1.err:
        return True
    return JsonResponse({"result": "bad", "mensaje": u"Problemas al ejecutar el reporte."})

def conviert_html_to_pdfsaveactas(template_src, context_dict, filename, folder):
    template = get_template(template_src)
    html = template.render(context_dict).encode(encoding="UTF-8")
    result = StringIO.BytesIO()
    output_folder = os.path.join(SITE_STORAGE, 'media', 'qrcode', folder)
    filepdf = open(output_folder + os.sep + filename, "w+b")
    links = lambda uri, rel: os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, ''))
    pdf1 = pisa.pisaDocument(StringIO.BytesIO(html), dest=filepdf, link_callback=links)
    pisaStatus = pisa.CreatePDF(StringIO.BytesIO(html), result, link_callback=links)
    if not pdf1.err:
        return True
    return JsonResponse({"result": "bad", "mensaje": u"Problemas al ejecutar el reporte."})

def conviert_html_to_pdfsavecontratomae(template_src, context_dict, filename, folder):
    template = get_template(template_src)
    html = template.render(context_dict).encode(encoding="UTF-8")
    result = StringIO.BytesIO()
    output_folder = os.path.join(SITE_STORAGE, 'media', 'qrcode', folder)
    filepdf = open(output_folder + os.sep + filename, "w+b")
    links = lambda uri, rel: os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, ''))
    pdf1 = pisa.pisaDocument(StringIO.BytesIO(html), dest=filepdf, link_callback=links)
    pisaStatus = pisa.CreatePDF(StringIO.BytesIO(html), result, link_callback=links)
    if not pdf1.err:
        return True
    return JsonResponse({"result": "bad", "mensaje": u"Problemas al ejecutar el reporte."})

def conviert_html_to_pdfsaveactagrado(template_src, context_dict, filename, folder):
    template = get_template(template_src)
    html = template.render(context_dict).encode(encoding="UTF-8")
    result = StringIO.BytesIO()
    output_folder = os.path.join(SITE_STORAGE, 'media', 'tematitulacionposgrado', folder)
    filepdf = open(output_folder + os.sep + filename, "w+b")
    links = lambda uri, rel: os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, ''))
    pdf1 = pisa.pisaDocument(StringIO.BytesIO(html), dest=filepdf, link_callback=links)
    pisaStatus = pisa.CreatePDF(StringIO.BytesIO(html), result, link_callback=links)
    if not pdf1.err:
        return True
    return JsonResponse({"result": "bad", "mensaje": u"Problemas al ejecutar el reporte."})

def conviert_html_to_pdfsaveqrguiapractica(template_src, context_dict, filename):
    template = get_template(template_src)
    html = template.render(context_dict).encode(encoding="UTF-8")
    result = StringIO.BytesIO()
    output_folder = os.path.join(SITE_STORAGE, 'media', 'qrcode', 'guiapractica')
    filepdf = open(output_folder + os.sep + filename, "w+b")
    links = lambda uri, rel: os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, ''))
    pdf1 = pisa.pisaDocument(StringIO.BytesIO(html), dest=filepdf, link_callback=links)
    pisaStatus = pisa.CreatePDF(StringIO.BytesIO(html), result, link_callback=links)
    if not pdf1.err:
        return True
    return JsonResponse({"result": "bad", "mensaje": u"Problemas al ejecutar el reporte."})

def conviert_html_to_pdfsaveinformeayudante(template_src, context_dict, filename):
    template = get_template(template_src)
    html = template.render(context_dict).encode(encoding="UTF-8")
    result = StringIO.BytesIO()
    output_folder = os.path.join(SITE_STORAGE, 'media', 'informeayudantecatedra')
    filepdf = open(output_folder + os.sep + filename, "w+b")
    links = lambda uri, rel: os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, ''))
    pdf1 = pisa.pisaDocument(StringIO.BytesIO(html), dest=filepdf, link_callback=links)
    pisaStatus = pisa.CreatePDF(StringIO.BytesIO(html), result, link_callback=links)
    if not pdf1.err:
        return HttpResponse(result.getvalue(), content_type='application/pdf')
    return JsonResponse({"result": "bad", "mensaje": u"Problemas al ejecutar el reporte."})

def conviert_html_to_pdf_parametros_save(template_src, context_dict, namefile, ruta1, ruta2):
    template = get_template(template_src)
    html = template.render(context_dict).encode(encoding="UTF-8")
    result = StringIO.BytesIO()
    output_folder = os.path.join(SITE_STORAGE, 'media', ruta1, ruta2)
    filepdf = open(output_folder + os.sep + "{}.pdf".format(namefile), "w+b")
    links = lambda uri, rel: os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, ''))
    pdf1 = pisa.pisaDocument(StringIO.BytesIO(html), dest=filepdf, link_callback=links)
    pisaStatus = pisa.CreatePDF(StringIO.BytesIO(html), result, link_callback=links)
    if not pdf1.err:
        # return HttpResponse(result.getvalue(), content_type='application/pdf')
        return True
    return JsonResponse({"result": "bad", "mensaje": u"Problemas al ejecutar el reporte."})

def write_pdf(template_src, context_dict, filename):
    template = get_template(template_src)
    output_folder = os.path.join(SITE_STORAGE, 'media', 'informespoa')
    try:
        os.makedirs(output_folder)
    except Exception as ex:
        pass
    context = Context(context_dict)
    # html = template.render(context)
    html = template.render(context).encode(encoding="UTF-8")
    filepdf = open(output_folder + os.sep + filename, "w+b")
    pdf = pisa.pisaDocument(StringIO.BytesIO(html), dest=filepdf, link_callback=fetch_resources)
    filepdf.close()
    return "".join(['informespoa', '/', filename])


def fetch_resources(uri, rel):
    return os.path.join(MEDIA_ROOT, uri.replace(MEDIA_URL, ""))


# -------------CREACION DE PDF IREPORTLAB--------------------------
documento = []
def generar_pdf_reportlab(topmargin = 90):
    from reportlab.lib.pagesizes import A4
    buff = BytesIO()
    response = HttpResponse(content_type='application/pdf')
    # --------DESCARGARLO DIRECTAMENTE---------------------
    # response['Content-Disposition'] = 'attachment; filename=%s' % nombre+".pdf"
    doc = SimpleDocTemplate(buff,
                            pagesize=A4,
                            # showBoundary=1,
                            leftMargin=25,
                            rightMargin=45,
                            topMargin=topmargin,
                            bottomMargin=27
                            )
    doc.build(documento,onFirstPage=addPageNumber, onLaterPages=addPageNumber)
    response.write(buff.getvalue())
    buff.close()
    return response


def addPageNumber(canvas, doc):
    from reportlab.lib.units import mm
    canvas.saveState()
    # ------------------------ENCABEZADO---------------------------------------------------------------------------
    header = Image(u'%s%s'%(settings.MEDIA_ROOT,'/reportes/encabezados_pies/cabecera_unemi.png'), width=555, height=75, hAlign='TOP')
    header.drawOn(canvas, 20, 755)
    # header = Image(u'%s%s' % (settings.MEDIA_ROOT, '/reportes/encabezados_pies/escudo.png'), width=80, height=75, hAlign='TOP')
    # header.drawOn(canvas, 20, 755)
    # header = Image(u'%s%s' % (settings.MEDIA_ROOT, '/reportes/encabezados_pies/logo.png'), width=120, height=75, hAlign='TOP')
    # header.drawOn(canvas, 475, 755)
    #----------------------------FOOTER--------------------------------------------------------------------------
    page_num = canvas.getPageNumber()
    text = "Página %s" % page_num
    canvas.setFont('Times-Roman', 9)
    canvas.drawRightString(200 * mm, 10 * mm, text)
    canvas.restoreState()


def add_tabla_reportlab(encabezado, detalles, anchocol,cabecera_left_center = [True, True], detalle_left_center = [False, True], combinar_celdas_horizontal=[],anchofila=None, tamano_letra_cabecera=9, tamano_letra_detalle=9, espacio=5):
    # cabecera_left_center y detalle_left_center.- Especificar cual contenidos de cada campos van a estar centrado o derecha, TRUE es centrado y FALSE es hacia la derecha
    # combinar_celdas_horizontal.- formato debe ser [(0,1)] ó [(0,1), (2,3)] etc..
    from reportlab.platypus import Table
    estilos = StyleSheet1()
    estilos.add(ParagraphStyle(name='Cabeceraleft', fontName='Helvetica-Bold', fontSize=tamano_letra_cabecera, leading=10, wordWrap=True, alignment=TA_LEFT, textColor=colors.black))
    estilos.add(ParagraphStyle(name='Cabeceracenter', fontName='Helvetica-Bold', fontSize=tamano_letra_cabecera, leading=10, wordWrap=True, alignment=TA_CENTER, textColor=colors.black))
    estilos.add(ParagraphStyle(name='Detallesleft', fontName='Helvetica', fontSize=tamano_letra_detalle, leading=10, wordWrap=True, alignment=TA_LEFT, textColor=colors.black))
    estilos.add(ParagraphStyle(name='Detallescenter', fontName='Helvetica', fontSize=tamano_letra_detalle, leading=10, wordWrap=True, alignment=TA_CENTER, textColor=colors.black))
    listaencabezado = []
    for row in encabezado:
        encabeza = []
        contando = 0
        for cell in row:
            if contando < cabecera_left_center.__len__():
                encabeza.append(Paragraph(cell.__str__(), estilos['Cabeceracenter'] if cabecera_left_center[contando] else estilos['Cabeceraleft']))
            else:
                encabeza.append(Paragraph(cell.__str__(), estilos['Cabeceraleft']))
            contando += 1
        listaencabezado.append(encabeza)
    listadetalles = []
    for rows in detalles:
        deta= []
        contando = 0
        for cell in rows:
            if contando < detalle_left_center.__len__():
                deta.append(Paragraph(cell.__str__(), estilos['Detallescenter'] if detalle_left_center[contando] else estilos['Detallesleft']))
            else:
                deta.append(Paragraph(cell.__str__(), estilos['Detallesleft']))
            contando += 1
        listadetalles.append(deta)
    t = Table(listaencabezado + listadetalles, colWidths=anchocol,rowHeights=anchofila, vAlign = 'MIDDLE', spaceAfter=espacio)
    estilos_tabla = [('GRID', (0, 0), (-1, -1), 1, colors.black), ('VALIGN', (0,0), (- 1, -1), 'MIDDLE')]
    for fila in combinar_celdas_horizontal:
        estilos_tabla.append(('SPAN', (fila[0], fila[1]),(fila[1], fila[0])))
    t.setStyle(TableStyle(estilos_tabla))
    documento.append(t)

def add_tabla_firma(detalles, anchocol,cabecera_left_center = [True, True], detalle_left_center = [False, True], combinar_celdas_horizontal=[],anchofila=None, tamano_letra_cabecera=9, tamano_letra_detalle=9):
    # cabecera_left_center y detalle_left_center.- Especificar cual contenidos de cada campos van a estar centrado o derecha, TRUE es centrado y FALSE es hacia la derecha
    # combinar_celdas_horizontal.- formato debe ser [(0,1)] ó [(0,1), (2,3)] etc..
    from reportlab.platypus import Table
    estilos = StyleSheet1()
    estilos.add(ParagraphStyle(name='Cabeceraleft', fontName='Helvetica-Bold', fontSize=tamano_letra_cabecera, leading=10, wordWrap=True, alignment=TA_LEFT, textColor=colors.black))
    estilos.add(ParagraphStyle(name='Cabeceracenter', fontName='Helvetica-Bold', fontSize=tamano_letra_cabecera, leading=10, wordWrap=True, alignment=TA_CENTER, textColor=colors.black))
    estilos.add(ParagraphStyle(name='Detallesleft', fontName='Helvetica', fontSize=tamano_letra_detalle, leading=10, wordWrap=True, alignment=TA_LEFT, textColor=colors.black))
    estilos.add(ParagraphStyle(name='Detallescenter', fontName='Helvetica', fontSize=tamano_letra_detalle, leading=10, wordWrap=True, alignment=TA_CENTER, textColor=colors.black))
    listadetalles = []
    for rows in detalles:
        deta= []
        contando = 0
        for cell in rows:
            if contando < detalle_left_center.__len__():
                deta.append(Paragraph(cell.__str__(), estilos['Detallescenter'] if detalle_left_center[contando] else estilos['Detallesleft']))
            else:
                deta.append(Paragraph(cell.__str__(), estilos['Detallesleft']))
            contando += 1
        listadetalles.append(deta)
    t = Table(listadetalles, colWidths=anchocol,rowHeights=anchofila, vAlign = 'MIDDLE', spaceAfter=100)
    estilos_tabla = [('GRID', (0, 0), (-1, -1), 1, colors.black), ('VALIGN', (0,0), (- 1, -1), 'BOTTOM')]
    for fila in combinar_celdas_horizontal:
        estilos_tabla.append(('SPAN', (fila[0], fila[1]),(fila[1], fila[0])))
    t.setStyle(TableStyle(estilos_tabla))
    documento.append(t)



def add_titulo_reportlab(descripcion, tamano = 12, alineacion = TA_CENTER, tipoletra = 'Helvetica-Bold', colortexto = colors.black, espacios = 19, afterespacio=0, beforeespacio=0):
    estilos = StyleSheet1()
    estilos.add(ParagraphStyle(name='Titulo', fontName=tipoletra, fontSize=tamano, leading=espacios, wordWrap=True, alignment=alineacion, spaceBefore= beforeespacio, spaceAfter=afterespacio, textColor=colortexto))
    header = Paragraph(descripcion, estilos['Titulo'])
    documento.append(header)


def add_graficos_barras_reportlab(datavalor, datanombres, minimo=None , maximo=None, step=None, anchografico=300, altografico=125, decimal=False, tamanoletra=6, posiciongrafico_x=50, posiciongrafico_y=30, colores = [], posicionleyenda_x = 450, titulo = 'Ningun titulo', tamanotitulo=10, tipoletratitulo = 'Helvetica-Bold', ubicaciontitulo_x = 0, ubicaciontitulo_y = 0, mostrarleyenda=True, barra_vertical_horizontal=False, presentar_nombre_o_numero=True):
    # formato de datavalor [(3,4,5)] o [[3,4,5]] o [(3,4,5), (6,7,8)] o [[3,4,5], [6,7,8]]
    # formato de datanombres [taller, laboratorio, practicas]
    # barra vertical=True y horizontal = False
    from reportlab.graphics.charts.barcharts import HorizontalBarChart
    from reportlab.graphics.charts.textlabels import Label
    from reportlab.lib.formatters import DecimalFormatter
    label = Label()
    label._text = titulo
    label.fontName = tipoletratitulo
    label.fontSize = tamanotitulo
    label.demo()
    label.textAnchor = 'middle'
    label.x = posiciongrafico_x + ubicaciontitulo_x
    label.y = posiciongrafico_y + altografico + ubicaciontitulo_y
    if presentar_nombre_o_numero:
        datossecuentes = [secuencia for secuencia in datanombres]
    else:
        datossecuentes = [secuencia.__str__() for secuencia in range(1, datanombres.__len__() + 1)]
    if not colores:
        colores = [colors.blue, colors.darkred, colors.greenyellow, colors.aqua, colors.beige, colors.aliceblue, colors.yellow, colors.antiquewhite]
    if minimo==None and maximo==None and step==None:
        minimo = 0
        maximo = max(datavalor)
        while not type(maximo) == int:
            if type(maximo) == float:
                maximo = int(maximo) + 1
            else:
                maximo = max(maximo)
        step = 2
        while not maximo%2==0:
            maximo += 1
        while not maximo %step==0 or maximo/step>12:
            step += 2
    d = Drawing(400, 200)
    if barra_vertical_horizontal:
        bc = VerticalBarChart()
    else:
        bc = HorizontalBarChart()
    bc.x = posiciongrafico_x
    bc.y = posiciongrafico_y
    bc.height = altografico
    bc.width = anchografico
    bc.data = datavalor
    bc.valueAxis.valueMin = minimo
    bc.valueAxis.valueMax = maximo
    bc.valueAxis.valueStep = step  # paso de distancia entre punto y punto
    bc.categoryAxis.labels.boxAnchor = 'ne'
    bc.valueAxis.labels.textAnchor = 'middle'
    if decimal:
        bc.barLabelFormat = DecimalFormatter(2, suffix='%')
    else:
        bc.barLabelFormat = '%d'
    bc.categoryAxis.labels.fontSize = tamanoletra
    bc.categoryAxis.categoryNames = datossecuentes
    bc.barSpacing = 0.5
    bc.barLabels.nudge = 7
    for i, color in enumerate(colores):
        bc.bars[i].fillColor = color
    colorynombres = []
    for i in range(0, datanombres.__len__()):
        colorynombres.append([colors.transparent, (datanombres[i].__str__()[0:20]+"..." if datanombres[i].__str__().__len__()>20 else datanombres[i].__str__(), (i + 1).__str__())])
    from reportlab.graphics.charts.legends import Legend
    legend = Legend()
    if mostrarleyenda:
        legend.x = posicionleyenda_x
        legend.y = altografico + posiciongrafico_y
        legend.dx = 8
        legend.dy = 8
        legend.fontName = 'Helvetica'
        legend.fontSize = 5
        legend.boxAnchor = 'n'
        legend.columnMaximum = 15
        legend.strokeWidth = 1
        legend.strokeColor = colors.black
        legend.deltax = 75
        legend.deltay = 10
        legend.autoXPadding = 5
        legend.yGap = 0
        legend.dxTextSpace = 5
        legend.alignment = 'right'
        legend.dividerLines = 1 | 2 | 4
        legend.dividerOffsY = 4.5
        legend.subCols.rpad = 10
        legend.colorNamePairs = colorynombres
    d.add(label)
    d.add(bc)
    if mostrarleyenda:
        d.add(legend)
    documento.append(d)


def add_graficos_circular_reporlab(datavalor, datanombres, anchografico = 150, altografico = 150, colores = [], posiciongrafico_x = 125, posiciongrafico_y = 30, posicionleyenda_x=450, titulo = 'Ningun titulo', tamanotitulo = 10, tipoletratitulo = 'Helvetica-Bold', ubicaciontitulo_x = 0, ubicaciontitulo_y = 0, mostrarsignoporcentaje = False, labelspie=[]):
    from xlwt.compat import xrange
    # pc.data = [10, 20, 30, 40, 50]
    # pc.labels = ['IE', 'Kopete', 'Chrome', 'Firefox', 'Opera']
    # formato de colores [colors.blue, colors.red, colors.green, colors.yellow, colors.beige, colors.aqua, colors.aliceblue, colors.antiquewhite]
    from reportlab.graphics.charts.piecharts import Pie
    from reportlab.graphics.charts.textlabels import Label
    label = Label()
    label._text = titulo
    label.fontName = tipoletratitulo
    label.fontSize = tamanotitulo
    label.textAnchor = 'middle'
    label.x = posiciongrafico_x + ubicaciontitulo_x
    label.y = posiciongrafico_y + altografico + ubicaciontitulo_y
    if not colores:
        colores = [colors.fidlightblue, colors.blue, colors.aliceblue, colors.antiquewhite, colors.green, colors.beige, colors.aqua, colors.red]
    d = Drawing(300, 200)
    pc = Pie()
    pc.x = posiciongrafico_x
    pc.y = posiciongrafico_y
    pc.width = anchografico
    pc.height = altografico
    pc.data = datavalor
    if labelspie:
        pc.sideLabels = 1
        pc.labels = labelspie
    pc.startAngle = 90
    pc.sameRadii = 1
    pc.slices.popout = 5
    for i, color in enumerate(colores):
        pc.slices[i].fillColor = color
    from reportlab.graphics.charts.legends import Legend
    legend = Legend()
    legend.x = posicionleyenda_x
    legend.y = altografico + posiciongrafico_y
    legend.dx = 8
    legend.dy = 8
    legend.fontName = 'Helvetica'
    legend.fontSize = 5
    legend.boxAnchor = 'n'
    legend.columnMaximum = 15
    legend.strokeWidth = 1
    legend.strokeColor = colors.black
    legend.deltax = 75
    legend.deltay = 10
    legend.autoXPadding = 5
    legend.yGap = 0
    legend.dxTextSpace = 5
    legend.alignment = 'right'
    legend.dividerLines = 1 | 2 | 4
    legend.dividerOffsY = 4.5
    legend.subCols.rpad = 10
    color_nombres = []
    for i in xrange(len(pc.data)):
        numero = '%0.2f' % pc.data[i]
        if mostrarsignoporcentaje:
            numero = numero.__str__()+"%"
        color_nombres.append([pc.slices[i].fillColor,(datanombres[i][0:20], numero)])
    legend.colorNamePairs = color_nombres
    d.add(label)
    d.add(pc)
    d.add(legend)
    documento.append(d)


def add_imagenes(url, ancho=350, alto=500, alineacion='CENTER'):
    imagen = Image(url, width=ancho, height=alto, hAlign=alineacion)
    documento.append(imagen)

def conviert_html_to_pdfsave_path(template_src, context_dict, output_folder, filename):
    try:
        css = open(os.path.join(SITE_ROOT, "static", "mimalla.css")).read()
        print(css)
        template = get_template(template_src)
        html = template.render(context_dict).encode(encoding="UTF-8")
        result = StringIO.BytesIO()
        filepdf = open(output_folder + os.sep + filename, "w+b")
        links = lambda uri, rel: os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, ''))
        pdf1 = pisa.pisaDocument(StringIO.BytesIO(html), dest=filepdf, default_css=css, link_callback=links)
        pisaStatus = pisa.CreatePDF(StringIO.BytesIO(html), result, link_callback=links)
        if not pdf1.err:
            return True, u'Pdf Generado Correctamente', filepdf.name
        else:
            NameError(str(pdf1.err))
    except Exception as ex:
        return False, u'Error: %s' % str(ex), None

def conviert_html_to_pdf_save_informe(template_src, context_dict, output_folder, filename):
    try:
        template = get_template(template_src)
        html = template.render(context_dict).encode(encoding="UTF-8")
        result = StringIO.BytesIO()
        output_folder = os.path.join(SITE_STORAGE, 'media', output_folder)
        filepdf = open(output_folder + os.sep + filename, "w+b")
        links = lambda uri, rel: os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, ''))
        pdf1 = pisa.pisaDocument(StringIO.BytesIO(html), dest=filepdf, link_callback=links)
        pisaStatus = pisa.CreatePDF(StringIO.BytesIO(html), dest=filepdf, link_callback=links)
        if not pisaStatus.err:
            return True, u'Pdf Generado Correctamente', filepdf.name
        else:
                NameError(str(pdf1.err))
    except Exception as ex:
        return False, u'Error: %s' % str(ex), None


def conviert_html_to_pdfsaveqrcertificado_generico(request, template_src, context_dict, output_folder, filename):
    template = get_template(template_src)
    html = template.render(context_dict, request).encode(encoding="UTF-8")
    result = StringIO.BytesIO()
    filepdf = open(output_folder + os.sep + filename, "w+b")
    links = lambda uri, rel: os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, ''))
    pdf1 = pisa.pisaDocument(StringIO.BytesIO(html), dest=filepdf, link_callback=link_callback)
    pisaStatus = pisa.CreatePDF(StringIO.BytesIO(html), result, link_callback=link_callback)
    if not pdf1.err:
        # return HttpResponse(result.getvalue(), content_type='application/pdf')
        return True, filepdf, result
    return False, pdf1, None


def conviert_html_to_pdfsaveqrcertificadoferiaparticipacion(request, template_src, context_dict, output_folder, filename):
    template = get_template(template_src)
    html = template.render(context_dict, request).encode(encoding="UTF-8")
    result = StringIO.BytesIO()
    filepdf = open(output_folder + os.sep + filename, "w+b")
    links = lambda uri, rel: os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, ''))
    #pdfmetrics.registerFont(TTFont('zhfont', os.path.join(SITE_ROOT, 'static/fonts/Poppins/Poppins-Regular.ttf')))
    pdfmetrics.registerFont(TTFont('zhfont', os.path.join(SITE_ROOT, 'static/fonts/Poppins/Poppins-Regular.ttf')))
    pdfmetrics.registerFont(TTFont('zhfont_bold', os.path.join(SITE_ROOT, 'static/fonts/Poppins/Poppins-Bold.ttf')))
    DEFAULT_FONT["helvetica"] = "zhfont"
    DEFAULT_FONT["helvetica-bold"] = "zhfont_bold"
    pdf1 = pisa.pisaDocument(StringIO.BytesIO(html), dest=filepdf, link_callback=links)
    pisaStatus = pisa.CreatePDF(StringIO.BytesIO(html), result, link_callback=links)
    if not pdf1.err:
        # return HttpResponse(result.getvalue(), content_type='application/pdf')
        return True, filepdf.name, result
    return False, pdf1, None


def conviert_html_to_pdfsaveqr_generico(request, template_src, context_dict, output_folder, filename):
    template = get_template(template_src)
    html = template.render(context_dict, request).encode(encoding="UTF-8")
    result = StringIO.BytesIO()
    filepdf = open(output_folder + os.sep + filename, "w+b")
    links = lambda uri, rel: os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, ''))
    pdf1 = pisa.pisaDocument(StringIO.BytesIO(html), dest=filepdf, link_callback=link_callback)
    #pisaStatus = pisa.CreatePDF(StringIO.BytesIO(html), result, link_callback=link_callback)
    if not pdf1.err:
        # return HttpResponse(result.getvalue(), content_type='application/pdf')
        return True, filepdf, result
    return False, pdf1, None


# App edcon anteproyecto
def conviert_html_to_pdfsave_anteproyecto(template_src, context_dict, filename, rutacarpeta):
    template = get_template(template_src)
    html = template.render(context_dict).encode(encoding="UTF-8")
    result = StringIO.BytesIO()
    output_folder = os.path.join(SITE_STORAGE, 'media', 'qrcode', rutacarpeta)
    filepdf = open(output_folder + os.sep + filename, "w+b")
    links = lambda uri, rel: os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, ''))
    pdf1 = pisa.pisaDocument(StringIO.BytesIO(html), dest=filepdf, link_callback=links)
    pisaStatus = pisa.CreatePDF(StringIO.BytesIO(html), result, link_callback=links)
    if not pdf1.err:
        return HttpResponse(result.getvalue(), content_type='application/pdf')
    return JsonResponse({"result": "bad", "mensaje": u"Problemas al ejecutar el reporte."})

def html_to_pdfsave_informemensualdocente(template_src, context_dict, filename, folder):
    template = get_template(template_src)
    html = template.render(context_dict).encode(encoding="UTF-8")
    result = StringIO.BytesIO()
    output_folder = os.path.join(SITE_STORAGE, 'media', folder)
    filepdf = open(output_folder + os.sep + filename, "w+b")
    links = lambda uri, rel: os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, ''))
    pdf1 = pisa.pisaDocument(StringIO.BytesIO(html), dest=filepdf, link_callback=links)
    pisaStatus = pisa.CreatePDF(StringIO.BytesIO(html), result, link_callback=links)
    if not pdf1.err:
        return True
    return JsonResponse({"result": "bad", "mensaje": u"Problemas al ejecutar el reporte."})

def conviert_html_to_pdfsave_generic_lotes(request, template_src, context_dict, output_folder, filename):
    try:
        if 'listado' in context_dict['data']:
            from sga.funciones import variable_valor
            varibale_global = variable_valor('LOTE_SIZE')
            lote_size = int(varibale_global) if varibale_global else 1000 # Tamaño de lote a procesar
            # lote_size = 5 # Tamaño de lote a procesar
            data = context_dict['data']['listado']
            context_dict['continit'] = 0
            if len(data) > lote_size:
                context=context_dict
                for i in range(0, len(data), lote_size):
                    lote_registros = data[i:i + lote_size]
                    template = get_template(template_src)
                    context['data']['listado'] = lote_registros
                    context['continit'] = i
                    html = template.render(context, request).encode(encoding="UTF-8")  # Generar HTML para el lote actual
                    filepdf = open(output_folder + os.sep + 'reporte_{}.pdf'.format(i // lote_size), "w+b")
                    # Convertir el HTML a PDF utilizando pisaDocument
                    pdf1=pisa.pisaDocument(html, dest=filepdf, link_callback=link_callback)
                    filepdf.close()
                    if pdf1.err:
                        return False
                merger = PdfFileMerger()

                for i in range(0, len(data), lote_size):
                    filename_1 = '{}\\reporte_{}.pdf'.format(output_folder, i // lote_size).replace("\\", '/')  # Nombre de archivo único para cada lote
                    merger.append(filename_1)

                output_filename = output_folder + '\\'+filename
                with open(output_filename.replace("\\", '/'), 'wb') as output:
                    merger.write(output)

                # Eliminar los archivos de lotes individuales
                # for i in range(0, len(data), lote_size):
                #     filename_2 = '{}\\reporte_{}.pdf'.format(output_folder, i // lote_size)
                #     os.remove(filename_2)
                return True
        template = get_template(template_src)
        html = template.render(context_dict, request).encode(encoding="UTF-8")
        filepdf = open(output_folder + os.sep + filename, "w+b")
        pdf1 = pisa.pisaDocument(StringIO.BytesIO(html), dest=filepdf, link_callback=link_callback)
        filepdf.close()
        if not pdf1.err:
            return True
        return False
    except Exception as ex:
        print(ex)
        raise NameError(f'Error: {ex}')


def conviert_html_to_pdf_titulacion_exa_complexivo(request, template_src, context_dict, output_folder, filename):
    template = get_template(template_src)
    html = template.render(context_dict, request).encode(encoding="UTF-8")
    result = StringIO.BytesIO()
    filepdf = open(output_folder + os.sep + filename, "w+b")
    links = lambda uri, rel: os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, ''))
    pdf1 = pisa.pisaDocument(StringIO.BytesIO(html), dest=filepdf, link_callback=link_callback)
    #pisaStatus = pisa.CreatePDF(StringIO.BytesIO(html), result, link_callback=link_callback)
    if not pdf1.err:
        # return HttpResponse(result.getvalue(), content_type='application/pdf')
        return True, filepdf, result
    return False, pdf1, None


def conviert_html_to_pdfsaveqr_cartaaceptacioncertificado(template_src, context_dict, filename):
    template = get_template(template_src)
    html = template.render(context_dict).encode(encoding="UTF-8")
    result = StringIO.BytesIO()
    output_folder = os.path.join(SITE_STORAGE, 'media', 'qrcode', 'cartaaceptacionposgrado')
    filepdf = open(output_folder + os.sep + filename, "w+b")
    links = lambda uri, rel: os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, ''))
    pdf1 = pisa.pisaDocument(StringIO.BytesIO(html), dest=filepdf, link_callback=links)
    pisaStatus = pisa.CreatePDF(StringIO.BytesIO(html), result, link_callback=links)
    pdf_file = ContentFile(result.getvalue())
    # Muestra el PDF en el navegador
    response = HttpResponse(result.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="documento.pdf"'
    if not pdf1.err:
            return True, pdf_file, response
    return JsonResponse({"result": "bad", "mensaje": u"Problemas al ejecutar el certificado."})

def html_to_pdfsave_evienciassalud(template_src, context_dict, filename, folder):
    template = get_template(template_src)
    html = template.render(context_dict).encode(encoding="UTF-8")
    result = StringIO.BytesIO()
    output_folder = os.path.join(SITE_STORAGE, 'media', folder)
    filepdf = open(output_folder + os.sep + filename, "w+b")
    links = lambda uri, rel: os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, ''))
    pdf1 = pisa.pisaDocument(StringIO.BytesIO(html), dest=filepdf, link_callback=links)
    pisaStatus = pisa.CreatePDF(StringIO.BytesIO(html), result, link_callback=links)
    if not pdf1.err and not pisaStatus.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename={}'.format(filename)
        return True, response
    return JsonResponse({"result": "bad", "mensaje": u"Problemas al ejecutar el reporte."})
