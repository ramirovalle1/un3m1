# -*- coding: UTF-8 -*-
import io
import json
import random
import sys
from datetime import datetime
import html
import re
from django.contrib.contenttypes.models import ContentType
from openpyxl import workbook as openxl
from openpyxl.styles import Font as openxlFont
from openpyxl.styles.alignment import Alignment as alin
import xlsxwriter
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Q
import xlwt
from django.forms import model_to_dict
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.template import Context
from django.template.loader import get_template
from django.views.decorators.csrf import csrf_exempt
from xlwt import *
from django.shortcuts import render

from balcon.models import RespuestaEncuestaSatisfaccion
from bd.models import IncidenciaSCRUM
from decorators import secure_module, last_access
from sagest.funciones import encrypt_id, encuesta_objeto
from sagest.models import SeccionDepartamento, Departamento
from settings import EMAIL_DOMAIN

from sga.commonviews import adduserdata

from sga.funciones import MiPaginador, log, generar_nombre, notificacion
from sga.models import Persona
from sga.templatetags.sga_extras import encrypt
from utils.filtros_genericos import filtro_persona_select, filtro_persona_select_v2
from .forms import RequerimientoPlanificacionAutomatizaForm
from .models import *
from django.db.models import Value, Count, Sum, F, FloatField
from django.db.models import Count, Case, When, CharField, F
from django.db.models.functions import Coalesce


@login_required(redirect_field_name='ret', login_url='/loginsagest')
@secure_module
@last_access
@transaction.atomic()
def view(request):
    data = {}
    adduserdata(request, data)
    persona = request.session['persona']
    usuario = request.user
    hoy = datetime.now()
    if request.method == 'POST':
        action = request.POST['action']

        if action == 'addrequerimiento':
            with transaction.atomic():
                try:
                    form = RequerimientoPlanificacionAutomatizaForm(request.POST)
                    idp = int(encrypt(request.POST['idp']))
                    if not form.is_valid():
                    # if not form.is_valid() or not form.validador(idp):
                        transaction.set_rollback(True)
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()], "mensaje": "Error en el formulario"})
                    requerimiento = RequerimientoPlanificacionAutomatiza(periodo_id=idp,
                                                                         gestion=form.cleaned_data['gestion'],
                                                                         prioridad=form.cleaned_data['prioridad'],
                                                                         tiporequerimiento=form.cleaned_data['tiporequerimiento'],
                                                                         # orden=form.cleaned_data['orden'],
                                                                         responsable=form.cleaned_data['responsable'],
                                                                         detalle=form.cleaned_data['detalle'],
                                                                         procedimiento=form.cleaned_data['procedimiento'],
                                                                         )
                    requerimiento.save(request)
                    documentos = request.FILES.getlist('adjuntos')
                    lista_items1 = json.loads(request.POST['lista_items1'])
                    for d in documentos:
                        items = [item for item in lista_items1 if item['archivo'] == d._name]
                        if len(items) > 1 and all(item['size'] == items[0]['size'] for item in items):
                            raise NameError(f'archivos duplicados {d._name}, remplace uno de los archivos duplicados.')
                        if d.size > 2194304:
                            raise NameError(f"archivo es mayor a 2 Mb. {items[0]['descripcion']}")
                        d._name = generar_nombre(f"adjunto_{requerimiento.id}_", d._name)
                        doc = DocumentoAdjuntoRequerimiento(requerimiento=requerimiento, leyenda=items[0]['descripcion'], archivo=d)
                        doc.save(request)

                    actividad = IncidenciaSCRUM(requerimiento=requerimiento,
                                                titulo=form.cleaned_data['procedimiento'],
                                                descripcion=form.cleaned_data['detalle'],
                                                prioridad=requerimiento.prioridad)
                    actividad.save(request)

                    for lider in actividad.lideres_departamento():
                        titulo = f'Incidencia recibida de {requerimiento.gestion.departamento}'
                        cuerpo = f'Estimado líder de equipo, se ingresó el requerimiento {actividad} de la unidad {requerimiento.gestion.departamento}, en caso de no pertenecer a su gestión hacer caso omiso '
                        notificacion(titulo, cuerpo, lider,
                                     None, '/adm_scrum_actividades?action=requerimientos', lider.pk, 2, 'sga-sagest',
                                     IncidenciaSCRUM, request)
                    log(u'Adicionó requerimiento: %s' % requerimiento, request, "add")
                    return JsonResponse({"result": False}, safe=False)
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": True, "mensaje": f'Su solicitud presenta el siguiente inconveniente: {ex.__str__()}'}, safe=False)

        if action == 'editrequerimiento':
            with transaction.atomic():
                try:
                    form = RequerimientoPlanificacionAutomatizaForm(request.POST)
                    id = int(encrypt(request.POST['id']))
                    idp = int(encrypt(request.POST['idp']))
                    if form.is_valid():
                        requerimiento = RequerimientoPlanificacionAutomatiza.objects.get(id=id)
                        requerimiento.gestion = form.cleaned_data['gestion']
                        requerimiento.prioridad = form.cleaned_data['prioridad']
                        requerimiento.tiporequerimiento = form.cleaned_data['tiporequerimiento']
                        requerimiento.responsable = form.cleaned_data['responsable']
                        requerimiento.detalle = form.cleaned_data['detalle']
                        requerimiento.procedimiento = form.cleaned_data['procedimiento']
                        requerimiento.save(request)

                        actividad = requerimiento.incidenciascrum_set.filter(status=True).first()
                        if actividad:
                            actividad.titulo = form.cleaned_data['procedimiento']
                            actividad.descripcion = form.cleaned_data['detalle']
                            actividad.prioridad = requerimiento.prioridad
                            actividad.save(request)

                        documentos = request.FILES.getlist('adjuntos')
                        lista_items1 = json.loads(request.POST['lista_items1'])
                        ids_excl = [int(item['id_adjunto']) for item in lista_items1 if item['id_adjunto']]
                        for d in documentos:
                            items = [item for item in lista_items1 if item['archivo'] == d._name]
                            if len(items) > 1 and all(item['size'] == items[0]['size'] for item in items):
                                raise NameError(f'Error, archivos duplicados {d._name}, remplace uno de los archivos duplicados.')
                            if d.size > 4194304:
                                raise NameError(f"Error, archivo es mayor a 4 Mb. {items[0]['descripcion']}")
                            if not items[0]['id_adjunto']:
                                d._name = generar_nombre(f"adjunto_{requerimiento.id}_", d._name)
                                doc = DocumentoAdjuntoRequerimiento(requerimiento=requerimiento, leyenda=items[0]['descripcion'], archivo=d)
                                doc.save(request)
                            else:
                                doc = DocumentoAdjuntoRequerimiento.objects.get(id=items[0]['id_adjunto'])
                                doc.leyenda = items[0]['descripcion']
                                doc.archivo = d
                                doc.save(request)
                            ids_excl.append(doc.id)
                        for items in lista_items1:
                            if items['id_adjunto']:
                                doc = DocumentoAdjuntoRequerimiento.objects.get(id=int(items['id_adjunto']))
                                doc.leyenda = items['descripcion']
                                doc.save(request, update_fields=['leyenda'])
                        requerimiento.documentos().exclude(id__in=ids_excl).update(status=False)
                        log(u'Edito requerimiento: %s' % requerimiento, request, "edit")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                             "mensaje": "Error en el formulario"})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        if action == 'delrequerimiento':
            with transaction.atomic():
                try:
                    instancia = RequerimientoPlanificacionAutomatiza.objects.get(pk=int(encrypt(request.POST['id'])))
                    instancia.status = False
                    instancia.save(request)
                    incidencias = instancia.incidenciascrum_set.filter(status=True)
                    incidencias.update(status=False)
                    log(u'Elimino requerimiento: %s' % instancia, request, "del")
                    res_json = {"error": False}
                except Exception as ex:
                    res_json = {'error': True, "message": "Error: {}".format(ex)}
                return JsonResponse(res_json, safe=False)

        # CALIFICAR ENCUESTA:
        elif action == 'calificarencuesta':
            try:
                instancia = RequerimientoPlanificacionAutomatiza.objects.get(id=encrypt_id(request.POST['id']))
                gestion = SeccionDepartamento.objects.get(id=encrypt_id(request.POST['idp']))
                content_type = ContentType.objects.get_for_model(instancia)
                preguntas = gestion.preguntas_encuesta()
                preguntasresueltas = json.loads(request.POST.get('lista_items1'))
                msg_error = 'Por favor complete la encuestas, marcando por lo menos una estrella en cada pregunta'
                if not len(preguntas) == len(preguntasresueltas):
                    raise NameError(msg_error)
                for pregunta in preguntasresueltas:
                    if int(pregunta['valoracion']) == 0:
                        raise NameError(msg_error)
                    respuesta = RespuestaEncuestaSatisfaccion(
                                pregunta_id=int(encrypt(pregunta['pregunta_id'])),
                                valoracion=int(pregunta['valoracion']),
                                observacion=pregunta['observacion'],
                                object_id=instancia.id,
                                content_type=content_type)
                    respuesta.save(request)
                    log(u'Calificó pregunta de modelo: %s' % instancia, request, "add")
                return JsonResponse({"result": False, "mensaje": "Se guardo correctamente la encuesta de satisfacción"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "%s" % ex.__str__()})

        elif action == 'evaluarservicio':
            try:
                acepta = request.POST['radio_select'] == 'acepta'
                realizo_encuesta = request.POST.get('realizo_encuesta', '')
                instancia = RequerimientoPlanificacionAutomatiza.objects.get(id=encrypt_id(request.POST['id']))
                if acepta:
                    instancia.estadoevaluacion = 2
                    instancia.save(request)
                    if realizo_encuesta:
                        gestion = SeccionDepartamento.objects.get(id=encrypt_id(request.POST['idp']))
                        content_type = ContentType.objects.get_for_model(instancia)
                        preguntas = gestion.preguntas_encuesta()
                        preguntasresueltas = json.loads(request.POST.get('lista_items1'))
                        msg_error = 'Por favor complete la encuestas, marcando por lo menos una estrella en cada pregunta'
                        if not len(preguntas) == len(preguntasresueltas):
                            raise NameError(msg_error)
                        for pregunta in preguntasresueltas:
                            if int(pregunta['valoracion']) == 0:
                                raise NameError(msg_error)
                            respuesta = RespuestaEncuestaSatisfaccion(
                                        pregunta_id=int(encrypt(pregunta['pregunta_id'])),
                                        valoracion=int(pregunta['valoracion']),
                                        observacion=pregunta['observacion'],
                                        object_id=instancia.id,
                                        content_type=content_type)
                            respuesta.save(request)
                            log(u'Calificó pregunta de modelo: %s' % instancia, request, "add")
                else:
                    observacion = request.POST.get('obs_rechazo', '')
                    if not observacion:
                        raise NameError('Por favor ingrese una observación')
                    instancia.observacionevaluacion = observacion
                    instancia.estadoevaluacion = 3
                    instancia.estado = 2
                    instancia.save(request)
                    actividad = instancia.incidencia()
                    actividad.estado = 2
                    actividad.save(request)
                    log(f'Cambio de estado a asignado por rechazo de evaluación: {instancia}', request, "edit")
                    titulo = f'Requerimiento rechazado por evaluación'
                    cuerpo = f'Estimado {actividad.asignadoa.nombre_completo_minus()}, el requerimiento {actividad} fue rechazado por el siguiente motivo: {observacion}'
                    notificacion(titulo, cuerpo, actividad.asignadoa,
                                 None, '/misactividades', actividad.asignadoa.pk, 2, 'sga-sagest',
                                 IncidenciaSCRUM, request)
                return JsonResponse({"result": False, "mensaje": "Datos guardados correctamente."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "%s" % ex.__str__()})

        return JsonResponse({"result": "bad", "mensaje": u"Solicitud Incorrecta."})
    else:
        if 'action' in request.GET:
            data['action'] = action = request.GET['action']
            idsdepa = persona.distributivopersona_set.values_list('unidadorganica_id', flat=True).filter(status=True,estadopuesto_id=1)
            idsreq = RequerimientoPlanificacionAutomatiza.objects.values_list('gestion__departamento_id',
                                                                     flat=True).filter(responsable=persona,
                                                                                       gestion__isnull=False,
                                                                                       status=True).order_by('-id')
            midepartamento_ = Departamento.objects.filter(Q(responsable=persona)|Q(responsable_subrogante=persona)
                                                           |Q(id__in=idsreq)
                                                           |Q(id__in=idsdepa),status=True, integrantes__isnull=False).distinct()

            if action == 'addrequerimiento':
                try:
                    data['idp'] = request.GET['id']
                    data['mi_departamento'] = midepartamento_.first()
                    req = RequerimientoPlanificacionAutomatiza.objects.filter(status=True, periodo_id=int(encrypt(request.GET['id'])),gestion__departamento__in=midepartamento_).order_by('orden').last()
                    cantidad = req.orden if req else 0
                    cantidad += 1
                    data['cantidad'] = cantidad
                    form = RequerimientoPlanificacionAutomatizaForm()
                    form.fields['responsable'].queryset = Persona.objects.filter(id=persona.id)
                    form.fields['responsable'].initial = Persona.objects.get(id=persona.id)
                    form.filtro_gestion(midepartamento_.values_list('id',flat=True))
                    data['form'] = form
                    data['seccionado'] = True
                    template = get_template("adm_ingresarequerimiento/modal/formrequerimiento.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'editrequerimiento':
                try:
                    data['filtro'] = requerimiento = RequerimientoPlanificacionAutomatiza.objects.get(pk=int(encrypt(request.GET['id'])))
                    data['idp'] = encrypt(requerimiento.periodo.id)
                    form = RequerimientoPlanificacionAutomatizaForm(initial=model_to_dict(requerimiento))
                    form.fields['responsable'].queryset = Persona.objects.filter(id=requerimiento.responsable.id)
                    form.filtro_gestion(midepartamento_.values_list('id',flat=True))
                    data['form'] = form
                    data['seccionado'] = True
                    template = get_template("adm_ingresarequerimiento/modal/formrequerimiento.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'requerimientos':
                try:
                    data['title'] = 'Requerimientos'
                    data['idp'] = idp = int(encrypt(request.GET['idp']))
                    data['plan'] = plan = PlanificacionAutomatiza.objects.get(id=idp)
                    url_vars, filtro, search, prioridad = f'&action={action}&idp={request.GET["idp"]}', \
                                                          Q(status=True, periodo_id=idp), \
                                                          request.GET.get('s', ''), request.GET.get('prioridad', '')
                    misreq = 'misreq' in request.GET
                    estado = request.GET.get('estado', '')

                    filtro = filtro & (Q(gestion__departamento__in=midepartamento_) | Q(responsable=persona))
                    if search:
                        data['s'] = search
                        filtro = filtro & (Q(procedimiento__icontains=search))
                        url_vars += f'&s={search}'

                    if prioridad:
                        data['prioridad'] = pr = int(prioridad)
                        filtro = filtro & (Q(prioridad=int(prioridad)))
                        url_vars += f'&prioridad={pr}'

                    if estado:
                        data['estado'] = estado = int(estado)
                        filtro = filtro & (Q(estado=estado))
                        url_vars += f'&estado={estado}'

                    if misreq:
                        filtro = filtro & (Q(responsable=persona))
                        url_vars += f'&misreq=1'
                        data['misreq'] = 1


                    planes = RequerimientoPlanificacionAutomatiza.objects.filter(filtro).order_by('-id')

                    paging = MiPaginador(planes, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['prioridades'] = PRIORIDAD
                    data['estados'] = ESTADO_REQUERIMIENTO
                    data['search'] = search if search else ""
                    data["url_vars"] = url_vars
                    data['listado'] = page.object_list
                    data['email_domain'] = EMAIL_DOMAIN
                    data['totalporevaluar'] = RequerimientoPlanificacionAutomatiza.objects.filter(status=True, periodo_id=idp, estado=3, estadoevaluacion=1, responsable=persona).count()
                    if 'exportar_excel' in request.GET:
                        pattern = re.compile('<.*?>')
                        wb = openxl.Workbook()
                        wb["Sheet"].title = "Reporte_requerimientos"
                        ws = wb.active
                        style_title = openxlFont(name='Arial', size=16, bold=True)
                        style_cab = openxlFont(name='Arial', size=10, bold=True)
                        alinear = alin(horizontal="center", vertical="center")
                        response = HttpResponse(content_type="application/ms-excel")
                        response[
                            'Content-Disposition'] = 'attachment; filename=Reporte de requerimientos' + '-' + random.randint(
                            1, 10000).__str__() + '.xlsx'
                        ws.column_dimensions['B'].width = 25
                        ws.column_dimensions['C'].width = 25
                        ws.column_dimensions['D'].width = 15
                        ws.column_dimensions['E'].width = 30
                        ws.column_dimensions['F'].width = 50
                        ws.column_dimensions['G'].width = 50
                        ws.column_dimensions['H'].width = 25
                        ws.column_dimensions['I'].width = 25
                        ws.column_dimensions['J'].width = 25
                        ws.column_dimensions['K'].width = 25
                        ws.column_dimensions['L'].width = 25
                        ws.merge_cells('A1:L1')
                        ws['A1'] = 'PLANIFICACIÓN DE REQUERIMIENTOS'
                        celda1 = ws['A1']
                        celda1.font = style_title
                        celda1.alignment = alinear

                        columns = [u"N°", u"PERIODO", u"GESTIÓN", u"PRIORIDAD", u"RESPONSABLE",
                                   u"DETALLE", u"PROCEDIMIENTO", u"FECHA SOLICITUD", u"FECHA ASIGNACIÓN", u"FECHA FINALIZACIÓN", u"USUARIO ASIGNADO"
                                   ]
                        row_num = 3
                        for col_num in range(0, len(columns)):
                            celda = ws.cell(row=row_num, column=(col_num + 1), value=columns[col_num])
                            celda.font = style_cab
                        row_num = 4
                        for list in planes:
                            ws.cell(row=row_num, column=1, value=list.orden)
                            ws.cell(row=row_num, column=2, value=str(list.periodo.nombre))
                            ws.cell(row=row_num, column=3, value=str(list.gestion))
                            ws.cell(row=row_num, column=4, value=str(list.get_prioridad_display()))
                            ws.cell(row=row_num, column=5, value=str(list.responsable.nombre_completo_minus()))
                            ws.cell(row=row_num, column=6, value=re.sub(pattern, '', html.unescape(list.detalle)))
                            ws.cell(row=row_num, column=7, value=re.sub(pattern, '', html.unescape(list.procedimiento)))
                            ws.cell(row=row_num, column=8, value=str(list.fecha_creacion.strftime("%Y-%m-%d")))
                            fasignacion, ffinalizacion, usuarioatendio = '', '', ''
                            if inc := list.incidencia():
                                fasignacion = inc.finicioactividad.strftime("%Y-%m-%d") if inc.finicioactividad else ''
                                ffinalizacion = inc.ffinactividad.strftime("%Y-%m-%d") if inc.ffinactividad else ''
                                usuarioatendio = inc.asignadoa.nombre_completo_minus() if inc.asignadoa else ''
                            ws.cell(row=row_num, column=9, value=str(fasignacion))
                            ws.cell(row=row_num, column=10, value=str(ffinalizacion))
                            ws.cell(row=row_num, column=11, value=str(usuarioatendio))
                            row_num += 1
                        wb.save(response)
                        return response
                    return render(request, 'adm_ingresarequerimiento/requerimientos.html', data)
                except Exception as ex:
                    pass

            if action == 'buscarpersonas':
                try:
                    id = request.GET.get('idsagregados', '')
                    integrantes = []
                    if id:
                        eDepartamento = Departamento.objects.get(id=int(id))
                        integrantes = eDepartamento.integrantes.all().values_list('id', flat=True)
                    resp = filtro_persona_select_v2(request, [], Q(id__in=integrantes))
                    return HttpResponse(json.dumps({'status': True, 'results': resp}))
                except Exception as ex:
                    pass

            if action == 'calificarencuesta':
                try:
                    eSeccion = SeccionDepartamento.objects.get(id=encrypt_id(request.GET['idex']))
                    data['encuesta'] = encuesta = encuesta_objeto(eSeccion).filter(vigente=True).first()
                    data['idp'] = eSeccion.id
                    data['id'] = encrypt_id(request.GET['id'])
                    template = get_template("alu_reservapolideportivo/modal/formencuesta.html")
                    json_content = template.render(data, request=request)
                    return JsonResponse({"result": True, 'data': json_content})
                except Exception as ex:
                    return JsonResponse({"result": False, "mensaje": u"%s" % ex.__str__()})

            if action == 'evaluarservicio':
                try:
                    eSeccion = SeccionDepartamento.objects.get(id=encrypt_id(request.GET['idex']))
                    data['idp'] = eSeccion.id
                    data['id'] = id_req = encrypt_id(request.GET['id'])
                    data['requerimiento'] = plan = RequerimientoPlanificacionAutomatiza.objects.get(id=id_req)
                    if not plan.respuestas_encuesta():
                        data['encuesta'] = encuesta_objeto(eSeccion).filter(vigente=True).first()
                    template = get_template("adm_ingresarequerimiento/modal/formraceptarorechazar.html")
                    json_content = template.render(data, request=request)
                    return JsonResponse({"result": True, 'data': json_content})
                except Exception as ex:
                    return JsonResponse({"result": False, "mensaje": u"%s" % ex.__str__()})

            if action == 'detallerequerimientoplanauto':
                try:
                    data['id'] = id = request.GET['id']
                    data['requerimiento'] = RequerimientoPlanificacionAutomatiza.objects.get(
                        pk=encrypt_id(request.GET['id']))
                    template = get_template("adm_ingresarequerimiento/modal/detallerequerimiento.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'message': str(ex)})

            return HttpResponseRedirect(request.path)
        else:
            data['title'] = u'Planificaciones'

            url_vars = ''
            filtro = Q(status=True)
            search = None
            ids = None
            tipo = None

            if 's' in request.GET:
                if request.GET['s'] != '':
                    search = request.GET['s']

            if search:
                filtro = filtro & (Q(nombre__icontains=search))
                url_vars += '&s=' + search

            planificacion = PlanificacionAutomatiza.objects.filter(filtro).order_by('-id')

            paging = MiPaginador(planificacion, 20)
            p = 1
            try:
                paginasesion = 1
                if 'paginador' in request.session:
                    paginasesion = int(request.session['paginador'])
                if 'page' in request.GET:
                    p = int(request.GET['page'])
                    p = int(request.GET['page'])
                else:
                    p = paginasesion
                try:
                    page = paging.page(p)
                except:
                    p = 1
                page = paging.page(p)
            except:
                page = paging.page(p)
            request.session['paginador'] = p
            data['paging'] = paging
            data['rangospaging'] = paging.rangos_paginado(p)
            data['page'] = page
            data['search'] = search if search else ""
            data["url_vars"] = url_vars
            data['ids'] = ids if ids else ""
            data['planificaciones'] = page.object_list
            data['email_domain'] = EMAIL_DOMAIN
            return render(request, 'adm_ingresarequerimiento/view.html', data)
