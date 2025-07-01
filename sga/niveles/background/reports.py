import json
import os
import io
import sys
import time
import threading
import random
import xlwt
import xlsxwriter
from datetime import datetime, timedelta
from django.contrib.auth.models import User
from django.core.exceptions import ObjectDoesNotExist
from django.db import connections, transaction
from django.db.models import Q, Sum
from xlwt import *
from openpyxl import load_workbook, workbook as openxl
from openpyxl.styles import Font as openxlFont
from openpyxl.styles.alignment import Alignment as alin
from webpush import send_user_notification
from sga.commonviews import traerNotificaciones
from settings import MEDIA_ROOT, MEDIA_URL, DEBUG


class reporte_conflicto_horario(threading.Thread):

    def __init__(self, request, data, notificacion_id):
        self.request = request
        self.data = data
        self.notificacion_id = notificacion_id
        threading.Thread.__init__(self)

    def run(self):
        from sga.commonviews import conflicto_materias_seleccionadas
        from sga.models import Nivel, Materia, Persona, Notificacion
        try:
            directory = os.path.join(MEDIA_ROOT, 'reportes', 'niveles')
            request, data, notificacion_id = self.request, self.data, self.notificacion_id
            usernotify = User.objects.get(pk=request.user.pk)
            ePersona = Persona.objects.get(usuario=usernotify)
            os.makedirs(directory, exist_ok=True)

            nombre_archivo = "conflicto_horario_materias{}.xls".format(random.randint(1, 10000).__str__())
            directory = os.path.join(MEDIA_ROOT, 'reportes', 'niveles', nombre_archivo)
            try:
                eNivel = Nivel.objects.get(pk=int(request.GET.get('idn', '0')))
            except ObjectDoesNotExist:
                raise NameError(u"No se encontro nivel académico")
            ePeriodo = eNivel.periodo

            __author__ = 'Unemi'
            title = easyxf('font: name Times New Roman, color-index black, bold on , height 350; alignment: horiz centre')
            subtitulos = easyxf('font: name Times New Roman, color-index black, bold on , height 200; alignment: horiz left')
            font_style = XFStyle()
            font_style.font.bold = True
            font_style2 = XFStyle()
            font_style2.font.bold = False
            wb = Workbook(encoding='utf-8')
            ws = wb.add_sheet('conflicto')
            ws.write_merge(0, 0, 0, 4, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
            ws.col(0).width = 7000
            ws.col(1).width = 7000
            ws.col(2).width = 7000
            ws.col(3).width = 1500
            ws.col(4).width = 10000
            ws.write(2, 0, 'PERIODO', subtitulos)
            ws.write(2, 1, 'FACULTAD', subtitulos)
            ws.write(2, 2, 'CARRERA', subtitulos)
            ws.write(2, 3, 'NIVEL', subtitulos)
            ws.write(2, 4, 'CONFLICTO', subtitulos)
            cursor = connections['sga_select'].cursor()
            sql = "select am.malla_id, am.nivelmalla_id , mat.paralelo from sga_materia mat, sga_nivel ni, sga_asignaturamalla am " \
                  " where mat.status=true and ni.status=true and mat.nivel_id=ni.id and am.status=true and mat.asignaturamalla_id=am.id " \
                  " and ni.periodo_id=" + str(ePeriodo.id) + " and ni.id=" + str(
                eNivel.id) + " group by am.malla_id,am.nivelmalla_id, mat.paralelo order by am.malla_id,am.nivelmalla_id, mat.paralelo"
            cursor.execute(sql)
            results = cursor.fetchall()
            a = 3
            for r in results:
                materias = Materia.objects.values_list('id', flat=True).filter(status=True, nivel__periodo=ePeriodo, asignaturamalla__malla__id=int(r[0]), asignaturamalla__nivelmalla__id=int(r[1]), paralelo=r[2])
                conflicto = conflicto_materias_seleccionadas(materias)
                if conflicto:
                    materia = Materia.objects.filter(status=True, pk=materias[0])[0]
                    ws.write(a, 0, u'%s' % ePeriodo.nombre)
                    ws.write(a, 1, u'%s' % materia.asignaturamalla.malla.carrera.coordinacion_set.filter(status=True)[0].nombre)
                    ws.write(a, 2, u'%s' % materia.asignaturamalla.malla.carrera.nombre)
                    ws.write(a, 3, u'%s' % r[1])
                    ws.write(a, 4, u'%s' % conflicto)
                    a += 1
            wb.save(directory)

            if notificacion_id > 0:
                eNotificacion = Notificacion.objects.get(pk=notificacion_id)
                eNotificacion.en_proceso = False
                eNotificacion.cuerpo = 'Generación de reporte de conflicto de horario listo'
                eNotificacion.url = "{}reportes/niveles/{}".format(MEDIA_URL, nombre_archivo)
                eNotificacion.save()
            else:
                eNotificacion = Notificacion(cuerpo='Generación de reporte de conflicto de horario listo',
                                             titulo='Reporte de conflicto de horario',
                                             destinatario=ePersona,
                                             url="{}reportes/niveles/{}".format(MEDIA_URL, nombre_archivo),
                                             prioridad=1,
                                             app_label='SGA',
                                             fecha_hora_visible=datetime.now() + timedelta(days=1),
                                             tipo=2,
                                             en_proceso=False)
                eNotificacion.save(request)
            try:
                send_user_notification(user=usernotify,
                                       payload={"head": "Reporte de conflicto de horario",
                                                "body": 'Generación de reporte de conflicto de horario listo',
                                                "action": "notificacion",
                                                "timestamp": time.mktime(datetime.now().timetuple()),
                                                "url": "{}reportes/niveles/{}".format(MEDIA_URL, nombre_archivo),
                                                "btn_notificaciones": traerNotificaciones(request, data, ePersona),
                                                "mensaje": 'Su reporte ha sido generado con exito'
                                                },
                                       ttl=500)
            except:
                pass
        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)
            if notificacion_id > 0:
                eNotificacion = Notificacion.objects.get(pk=notificacion_id)
                eNotificacion.en_proceso = False
                eNotificacion.error = True
                eNotificacion.titulo = 'Reporte de conflicto de horario falló en la ejecución'
                eNotificacion.cuerpo = textoerror
                eNotificacion.save()
            else:
                eNotificacion = Notificacion(cuerpo=textoerror,
                                             titulo='Reporte de conflicto de horario falló en la ejecución',
                                             destinatario=ePersona,
                                             prioridad=1,
                                             app_label='SGA',
                                             fecha_hora_visible=datetime.now() + timedelta(days=1),
                                             tipo=2,
                                             en_proceso=False,
                                             error=True)
                eNotificacion.save(request)

            try:
                send_user_notification(user=usernotify,
                                       payload={"head": "Reporte de conflicto de horario Fallido",
                                                "body": 'Reporte de Planificación de Matriculas a fallado',
                                                "action": "notificacion",
                                                "timestamp": time.mktime(datetime.now().timetuple()),
                                                "btn_notificaciones": traerNotificaciones(request, data, ePersona),
                                                "mensaje": textoerror,
                                                "error": True
                                                },
                                       ttl=500)
            except:
                pass

class reporte_situacion_academica_coordinacion(threading.Thread):

    def __init__(self, request, data, notificacion_id):
        self.request = request
        self.data = data
        self.notificacion_id = notificacion_id
        threading.Thread.__init__(self)

    def run(self):
        from sga.models import Coordinacion, Periodo, Persona, Notificacion
        try:
            directory = os.path.join(MEDIA_ROOT, 'reportes', 'niveles')
            request, data, notificacion_id = self.request, self.data, self.notificacion_id
            usernotify = User.objects.get(pk=request.user.pk)
            ePersona = Persona.objects.get(usuario=usernotify)
            os.makedirs(directory, exist_ok=True)
            try:
                ePeriodo = Periodo.objects.get(pk=int(request.GET.get('idp', '0')))
            except ObjectDoesNotExist:
                raise NameError(u"No se encontro coordinacion académica")
            try:
                eCoordinacion = Coordinacion.objects.get(pk=int(request.GET.get('idc', '0')))
            except ObjectDoesNotExist:
                raise NameError(u"No se encontro coordinacion académica")
            nombre_archivo = "situacion_academica_coordinacion_{}_{}.xls".format(eCoordinacion.alias.lower() if eCoordinacion.alias else eCoordinacion.nombre.lower(), random.randint(1, 10000).__str__())
            directory = os.path.join(MEDIA_ROOT, 'reportes', 'niveles', nombre_archivo)

            __author__ = 'UNEMI'
            title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
            font_style = XFStyle()
            font_style.font.bold = True
            font_style2 = XFStyle()
            font_style2.font.bold = False
            wb = Workbook(encoding='utf-8')
            ws = wb.add_sheet('Hoja1')
            ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
            columns = [
                (u"PERIODO", 6000),
                (u"FACULTAD", 6000),
                (u"CARRERA", 6000),
                (u"NIVEL", 6000),
                (u"PARALELO", 6000),
                (u"DOCENTE", 6000),
                (u"ASIGNATURA", 6000),
                (u"PROMEDIO PARCIAL 1", 6000),
                (u"PROMEDIO PARCIAL 1 CON NOTA", 6000),
                (u"PROMEDIO PARCIAL 2", 6000),
                (u"PROMEDIO PARCIAL 2 CON NOTA", 6000),
                (u"PROMEDIO NOTA FINAL", 6000),
                (u"PROMEDIO NOTA FINAL CON NOTA", 6000),
            ]
            row_num = 3
            for col_num in range(len(columns)):
                ws.write(row_num, col_num, columns[col_num][0], font_style)
                ws.col(col_num).width = columns[col_num][1]
            cursor = connections['sga_select'].cursor()
            sql = "SELECT DISTINCT sga_periodo.nombre AS Periodo, sga_coordinacion.nombre AS Facultad, sga_carrera.nombre AS Carrera, " \
                  " sga_nivelmalla.nombre as Nivel, sga_materia.paralelo as Paralelo, (sga_persona.apellido1 || ' ' || sga_persona.apellido2 || ' ' || sga_persona.nombres) AS Docente, " \
                  " sga_asignatura.nombre AS Asignatura, (select round(avg(valor),2) from public.sga_materiaasignada sga_materiaasignada, public.sga_matricula sga_matricula , public.sga_evaluaciongenerica sga_evaluaciongenerica, public.sga_detallemodeloevaluativo sga_detallemodeloevaluativo where sga_materiaasignada.matricula_id=sga_matricula.id and sga_matricula.estado_matricula in (2,3) " \
                  " and sga_materiaasignada.materia_id=sga_materia.id and sga_evaluaciongenerica.materiaasignada_id=sga_materiaasignada.id and sga_detallemodeloevaluativo.id=sga_evaluaciongenerica.detallemodeloevaluativo_id and sga_detallemodeloevaluativo.id in (14,4)) as P1COMPLETO, " \
                  " (select round(avg(valor),2) from public.sga_materiaasignada sga_materiaasignada, public.sga_matricula sga_matricula ,public.sga_evaluaciongenerica sga_evaluaciongenerica, public.sga_detallemodeloevaluativo sga_detallemodeloevaluativo where sga_materiaasignada.matricula_id=sga_matricula.id and sga_matricula.estado_matricula in (2,3) and sga_materiaasignada.materia_id=sga_materia.id " \
                  " and sga_evaluaciongenerica.materiaasignada_id=sga_materiaasignada.id and sga_detallemodeloevaluativo.id=sga_evaluaciongenerica.detallemodeloevaluativo_id and sga_detallemodeloevaluativo.id in (14,4) and sga_materiaasignada.notafinal>0) as P1CONNOTA, " \
                  " (select round(avg(valor),2) from public.sga_materiaasignada sga_materiaasignada,  public.sga_matricula sga_matricula, public.sga_evaluaciongenerica sga_evaluaciongenerica, public.sga_detallemodeloevaluativo sga_detallemodeloevaluativo where sga_materiaasignada.matricula_id=sga_matricula.id and sga_matricula.estado_matricula in (2,3) and sga_materiaasignada.materia_id=sga_materia.id " \
                  " and sga_evaluaciongenerica.materiaasignada_id=sga_materiaasignada.id and sga_detallemodeloevaluativo.id=sga_evaluaciongenerica.detallemodeloevaluativo_id and sga_detallemodeloevaluativo.nombre='P2') as P2COMPLETO, " \
                  " (select round(avg(valor),2) from public.sga_materiaasignada sga_materiaasignada, public.sga_matricula sga_matricula ,public.sga_evaluaciongenerica sga_evaluaciongenerica, public.sga_detallemodeloevaluativo sga_detallemodeloevaluativo where sga_materiaasignada.matricula_id=sga_matricula.id and sga_matricula.estado_matricula in (2,3) and sga_materiaasignada.materia_id=sga_materia.id " \
                  " and sga_evaluaciongenerica.materiaasignada_id=sga_materiaasignada.id and sga_detallemodeloevaluativo.id=sga_evaluaciongenerica.detallemodeloevaluativo_id and sga_detallemodeloevaluativo.nombre='P2'  and sga_materiaasignada.notafinal>0) as P2CONNOTA, " \
                  " (select round(avg(m1.notafinal),2) from sga_materiaasignada m1, sga_matricula mat where m1.matricula_id=mat.id and mat.estado_matricula in (2,3) and m1.materia_id=sga_materia.id and m1.notafinal>0) as Nota_FinalCONNOTA, " \
                  " (select round(avg(m1.notafinal),2) from sga_materiaasignada m1, sga_matricula mat where m1.matricula_id=mat.id and mat.estado_matricula in (2,3) and m1.materia_id=sga_materia.id) as Nota_FinalCOMPLETO FROM public.sga_profesor sga_profesor RIGHT OUTER JOIN public.sga_profesormateria sga_profesormateria ON sga_profesor.id = sga_profesormateria.profesor_id inner JOIN public.sga_persona sga_persona ON sga_profesor.persona_id = sga_persona.id inner JOIN public.sga_materia sga_materia ON sga_profesormateria.materia_id = sga_materia.id inner JOIN public.sga_nivel sga_nivel ON sga_materia.nivel_id = sga_nivel.id inner JOIN public.sga_asignatura sga_asignatura ON sga_materia.asignatura_id = sga_asignatura.id inner JOIN public.sga_asignaturamalla sga_asignaturamalla ON sga_materia.asignaturamalla_id = sga_asignaturamalla.id inner join public.sga_nivelmalla sga_nivelmalla ON sga_nivelmalla.id=sga_asignaturamalla.nivelmalla_id inner JOIN public.sga_malla sga_malla ON sga_asignaturamalla.malla_id = sga_malla.id inner JOIN public.sga_carrera sga_carrera ON sga_malla.carrera_id = sga_carrera.id inner JOIN public.sga_coordinacion_carrera sga_coordinacion_carrera ON sga_carrera.id = sga_coordinacion_carrera.carrera_id inner JOIN public.sga_coordinacion sga_coordinacion ON sga_coordinacion_carrera.coordinacion_id = sga_coordinacion.id inner JOIN public.sga_periodo sga_periodo ON sga_nivel.periodo_id = sga_periodo.id WHERE sga_profesormateria.principal = True And sga_periodo.id = '" + str(ePeriodo.id) + "' And sga_coordinacion.id = '" + str(eCoordinacion.id) + "' ORDER BY sga_carrera.nombre ASC, Docente ASC, sga_asignatura.nombre ASC"
            cursor.execute(sql)
            results = cursor.fetchall()
            row_num = 4
            for r in results:
                i = 0
                campo1 = r[0]
                campo2 = r[1]
                campo3 = r[2]
                campo4 = r[3]
                campo5 = r[4]
                campo6 = r[5]
                campo7 = r[6]
                campo8 = r[7]
                campo9 = r[8]
                campo10 = r[9]
                campo11 = r[10]
                campo12 = r[12]
                campo13 = r[11]
                ws.write(row_num, 0, campo1, font_style2)
                ws.write(row_num, 1, campo2, font_style2)
                ws.write(row_num, 2, campo3, font_style2)
                ws.write(row_num, 3, campo4, font_style2)
                ws.write(row_num, 4, campo5, font_style2)
                ws.write(row_num, 5, campo6, font_style2)
                ws.write(row_num, 6, campo7, font_style2)
                ws.write(row_num, 7, campo8, font_style2)
                ws.write(row_num, 8, campo9, font_style2)
                ws.write(row_num, 9, campo10, font_style2)
                ws.write(row_num, 10, campo11, font_style2)
                ws.write(row_num, 11, campo12, font_style2)
                ws.write(row_num, 12, campo13, font_style2)
                row_num += 1
            wb.save(directory)

            if notificacion_id > 0:
                eNotificacion = Notificacion.objects.get(pk=notificacion_id)
                eNotificacion.en_proceso = False
                eNotificacion.cuerpo = f'Reporte de situación académica {eCoordinacion.nombre.lower()} listo'
                eNotificacion.url = "{}reportes/niveles/{}".format(MEDIA_URL, nombre_archivo)
                eNotificacion.save()
            else:
                eNotificacion = Notificacion(cuerpo='Generación de reporte de situación académica listo',
                                             titulo='Reporte de situación académica',
                                             destinatario=ePersona,
                                             url="{}reportes/niveles/{}".format(MEDIA_URL, nombre_archivo),
                                             prioridad=1,
                                             app_label='SGA',
                                             fecha_hora_visible=datetime.now() + timedelta(days=1),
                                             tipo=2,
                                             en_proceso=False)
                eNotificacion.save(request)
            try:
                send_user_notification(user=usernotify,
                                       payload={"head": "Reporte de situación académica",
                                                "body": 'Generación de reporte de situación académica listo',
                                                "action": "notificacion",
                                                "timestamp": time.mktime(datetime.now().timetuple()),
                                                "url": "{}reportes/niveles/{}".format(MEDIA_URL, nombre_archivo),
                                                "btn_notificaciones": traerNotificaciones(request, data, ePersona),
                                                "mensaje": 'Su reporte ha sido generado con exito'
                                                },
                                       ttl=500)
            except:
                pass
        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)
            if notificacion_id > 0:
                eNotificacion = Notificacion.objects.get(pk=notificacion_id)
                eNotificacion.en_proceso = False
                eNotificacion.error = True
                eNotificacion.titulo = f'Reporte de situación académica {eCoordinacion.nombre.lower()} falló en la ejecución'
                eNotificacion.cuerpo = textoerror
                eNotificacion.save()
            else:
                eNotificacion = Notificacion(cuerpo=textoerror,
                                             titulo='Reporte de situación académica falló en la ejecución',
                                             destinatario=ePersona,
                                             prioridad=1,
                                             app_label='SGA',
                                             fecha_hora_visible=datetime.now() + timedelta(days=1),
                                             tipo=2,
                                             en_proceso=False,
                                             error=True)
                eNotificacion.save(request)

            try:
                send_user_notification(user=usernotify,
                                       payload={"head": "Reporte de situación académica Fallido",
                                                "body": 'Generación del reporte de situación académica a fallado',
                                                "action": "notificacion",
                                                "timestamp": time.mktime(datetime.now().timetuple()),
                                                "btn_notificaciones": traerNotificaciones(request, data, ePersona),
                                                "mensaje": textoerror,
                                                "error": True
                                                },
                                       ttl=500)
            except:
                pass


class reporte_estudiante_limpio_coordinacion(threading.Thread):

    def __init__(self, request, data, notificacion_id):
        self.request = request
        self.data = data
        self.notificacion_id = notificacion_id
        threading.Thread.__init__(self)

    def run(self):
        from sga.models import Coordinacion, Periodo, Persona, Notificacion, Inscripcion, AsignaturaMalla, \
            RecordAcademico
        try:
            directory = os.path.join(MEDIA_ROOT, 'reportes', 'niveles')
            request, data, notificacion_id = self.request, self.data, self.notificacion_id
            usernotify = User.objects.get(pk=request.user.pk)
            ePersona = Persona.objects.get(usuario=usernotify)
            os.makedirs(directory, exist_ok=True)

            try:
                ePeriodo = Periodo.objects.get(pk=int(request.GET.get('idp', '0')))
            except ObjectDoesNotExist:
                raise NameError(u"No se encontro coordinacion académica")
            try:
                eCoordinacion = Coordinacion.objects.get(pk=int(request.GET.get('idc', '0')))
            except ObjectDoesNotExist:
                raise NameError(u"No se encontro coordinacion académica")

            nombre_archivo = "estudiantes_limpio_coordinacion_{}_{}.xls".format(eCoordinacion.alias.lower() if eCoordinacion.alias else eCoordinacion.nombre.lower(), random.randint(1, 10000).__str__())
            directory = os.path.join(MEDIA_ROOT, 'reportes', 'niveles', nombre_archivo)
            __author__ = 'UNEMI'
            font_style = XFStyle()
            font_style.font.bold = True
            font_style2 = XFStyle()
            font_style2.font.bold = False
            wb = Workbook(encoding='utf-8')
            ws = wb.add_sheet('Hoja1')
            ws.col(0).width = 6000
            ws.col(1).width = 6000
            ws.col(2).width = 6000
            ws.col(3).width = 6000
            ws.col(4).width = 6000
            ws.write(0, 0, 'CEDULA')
            ws.write(0, 1, 'CARRERA_ID')
            ws.write(0, 2, 'SECCION_ID')
            ws.write(0, 3, 'PARALELO')
            ws.write(0, 4, 'NIVEL_MATRICULA')

            a = 0
            date_format = xlwt.XFStyle()
            date_format.num_format_str = 'yyyy/mm/dd'
            if not DEBUG:
                for eInscripcion in Inscripcion.objects.filter(matricula__nivel__periodo=ePeriodo, carrera__coordinacion=eCoordinacion).exclude(coordinacion_id=9).order_by('coordinacion', 'carrera'):
                    malla = eInscripcion.carrera.malla()
                    nivelmalla_matricula = eInscripcion.matricula_periodo(ePeriodo).nivelmalla_id
                    asignaturas_atras = AsignaturaMalla.objects.filter(nivelmalla__id__lte=nivelmalla_matricula, malla=malla)
                    asignaturas_delante = AsignaturaMalla.objects.filter(nivelmalla__id__gt=nivelmalla_matricula, malla=malla)
                    bandera = 0
                    for asignatura_atras in asignaturas_atras:
                        if bandera == 0:
                            if (eRecordAcademico := RecordAcademico.objects.filter(inscripcion=eInscripcion, asignatura=asignatura_atras.asignatura).first()) is not None:
                                if eRecordAcademico.aprobada == False:
                                    bandera = 1
                            else:
                                bandera = 1
                    if bandera == 0:
                        for asignatura_delante in asignaturas_delante:
                            if RecordAcademico.objects.filter(inscripcion=eInscripcion, asignatura=asignatura_delante.asignatura).exists():
                                bandera = 1
                    if bandera == 0:
                        a += 1
                        ws.write(a, 0, u'%s' % eInscripcion.persona.cedula)
                        ws.write(a, 1, u'%s' % malla.carrera_id)
                        ws.write(a, 2, u'%s' % eInscripcion.sesion_id)
                        ws.write(a, 3, '')
                        ws.write(a, 4, u'%s' % nivelmalla_matricula)

            wb.save(directory)

            if notificacion_id > 0:
                eNotificacion = Notificacion.objects.get(pk=notificacion_id)
                eNotificacion.en_proceso = False
                eNotificacion.cuerpo = f'Generación del reporte de estudiante limpio {eCoordinacion.nombre.lower()} listo'
                eNotificacion.url = "{}reportes/niveles/{}".format(MEDIA_URL, nombre_archivo)
                eNotificacion.save()
            else:
                eNotificacion = Notificacion(cuerpo=f'Generación del reporte de estudiante limpio {eCoordinacion.nombre.lower()} listo',
                                             titulo=f'Reporte de estudiante limpio {eCoordinacion.alias if eCoordinacion.alias else eCoordinacion.nombre}',
                                             destinatario=ePersona,
                                             url="{}reportes/niveles/{}".format(MEDIA_URL, nombre_archivo),
                                             prioridad=1,
                                             app_label='SGA',
                                             fecha_hora_visible=datetime.now() + timedelta(days=1),
                                             tipo=2,
                                             en_proceso=False)
                eNotificacion.save(request)
            try:
                send_user_notification(user=usernotify,
                                       payload={"head": "Reporte de estudiante limpio",
                                                "body": 'Generación de reporte de estudiante limpio listo',
                                                "action": "notificacion",
                                                "timestamp": time.mktime(datetime.now().timetuple()),
                                                "url": "{}reportes/niveles/{}".format(MEDIA_URL, nombre_archivo),
                                                "btn_notificaciones": traerNotificaciones(request, data, ePersona),
                                                "mensaje": 'Su reporte ha sido generado con exito'
                                                },
                                       ttl=500)
            except:
                pass
        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)
            if notificacion_id > 0:
                eNotificacion = Notificacion.objects.get(pk=notificacion_id)
                eNotificacion.en_proceso = False
                eNotificacion.error = True
                eNotificacion.titulo = f'Reporte de estudiante limpio {eCoordinacion.alias if eCoordinacion.alias else eCoordinacion.nombre} falló en la ejecución'
                eNotificacion.cuerpo = textoerror
                eNotificacion.save()
            else:
                eNotificacion = Notificacion(cuerpo=textoerror,
                                             titulo=f'Reporte de estudiante limpio {eCoordinacion.alias if eCoordinacion.alias else eCoordinacion.nombre} falló en la ejecución',
                                             destinatario=ePersona,
                                             prioridad=1,
                                             app_label='SGA',
                                             fecha_hora_visible=datetime.now() + timedelta(days=1),
                                             tipo=2,
                                             en_proceso=False,
                                             error=True)
                eNotificacion.save(request)
            try:
                send_user_notification(user=usernotify,
                                       payload={"head": "Reporte de estudiante limpio Fallido",
                                                "body": 'Generación del reporte de estudiante limpio a fallado',
                                                "action": "notificacion",
                                                "timestamp": time.mktime(datetime.now().timetuple()),
                                                "btn_notificaciones": traerNotificaciones(request, data, ePersona),
                                                "mensaje": textoerror,
                                                "error": True
                                                },
                                       ttl=500)
            except:
                pass


class reporte_resultados_nivelacion(threading.Thread):

    def __init__(self, request, data, notificacion_id):
        self.request = request
        self.data = data
        self.notificacion_id = notificacion_id
        threading.Thread.__init__(self)

    def run(self):
        from sga.models import Coordinacion, Periodo, Persona, Notificacion
        try:
            directory = os.path.join(MEDIA_ROOT, 'reportes', 'niveles')
            request, data, notificacion_id = self.request, self.data, self.notificacion_id
            usernotify = User.objects.get(pk=request.user.pk)
            ePersona = Persona.objects.get(usuario=usernotify)
            os.makedirs(directory, exist_ok=True)

            try:
                ePeriodo = Periodo.objects.get(pk=int(request.GET.get('idp', '0')))
            except ObjectDoesNotExist:
                raise NameError(u"No se encontro coordinacion académica")
            try:
                eCoordinacion = Coordinacion.objects.get(pk=int(request.GET.get('idc', '0')))
            except ObjectDoesNotExist:
                raise NameError(u"No se encontro coordinacion académica")

            nombre_archivo = "reporte_resultados_finales_nivelacion_{}.xlsx".format(random.randint(1, 10000).__str__())
            directory = os.path.join(MEDIA_ROOT, 'reportes', 'niveles', nombre_archivo)
            workbook = xlsxwriter.Workbook(directory, {'constant_memory': True})
            ws = workbook.add_worksheet('Hoja1')
            merge_format = workbook.add_format({'bold': 1,
                                                'border': 1,
                                                'align': 'center',
                                                'valign': 'vcenter',
                                                'bg_color': 'silver',
                                                'text_wrap': 1})
            formatocelda = workbook.add_format({'border': 1})
            ws.merge_range(0, 0, 0, 11, 'UNIVERSIDAD ESTATAL DE MILAGRO', workbook.add_format({'align': 'center', 'valign': 'vcenter', 'bold': 1, 'font_size': 16}))
            columns = [(u"CÉDULA", 20),
                       (u"APELLIDOS", 40),
                       (u"NOMBRES", 40),
                       (u"CARRERA", 40),
                       (u"MODALIDAD", 20),
                       (u"ESTATUS", 20),
                       (u"PROMEDIO FINAL", 15),
                       (u"MATERIA 1", 35),
                       (u"MATERIA 2", 35),
                       (u"MATERIA 3", 35),
                       (u"MATERIA 4", 35),
                       (u"NÚMERO DE MATRÍCULA", 15)
                       ]
            row_num = 3
            for col_num in range(len(columns)):
                ws.write(row_num, col_num, columns[col_num][0], merge_format)
                ws.set_column(col_num, col_num, columns[col_num][1])

            cursor = connections['default'].cursor()
            sql = """select p.cedula AS cedula, p.apellido1||' '||p.apellido2 AS apellidos, p.nombres AS nombres, car.nombre as carrera,  
                                        mo.nombre as modalidad,Case When mat.aprobado=TRUE THEN 'APROBADO' ELSE 'REPROBADO' End AS status,
                                        (select avg(notafinal)
                                        from sga_matricula m
                                        inner join sga_materiaasignada ma on ma.matricula_id=m.id
                                        where m.id=mat.id) as promedio,
                                        (select ma.matriculas
                                        from sga_matricula matri
                                        inner join sga_materiaasignada ma on ma.matricula_id=matri.id
                                        where matri.id=mat.id and ma.matriculas=1 order by ma.id limit 1) primeravez,
                                        (select ma.matriculas
                                        from sga_matricula matri
                                        inner join sga_materiaasignada ma on ma.matricula_id=matri.id
                                        where matri.id=mat.id and ma.matriculas>1 order by ma.id limit 1) segundavez,
                                        (select a.nombre
                                        from sga_matricula matri
                                        inner join sga_materiaasignada ma on ma.matricula_id=matri.id
                                        inner join sga_materia mate on mate.id=ma.materia_id
                                        inner join sga_asignatura a on a.id=mate.asignatura_id
                                        where matri.id=mat.id order by ma.id  limit 1 offset 0) as asignatura1,
                                        (select a.nombre
                                        from sga_matricula matri
                                        inner join sga_materiaasignada ma on ma.matricula_id=matri.id
                                        inner join sga_materia mate on mate.id=ma.materia_id
                                        inner join sga_asignatura a on a.id=mate.asignatura_id
                                        where matri.id=mat.id order by ma.id  limit 1 offset 1) as asignatura2,
                                        (select a.nombre
                                        from sga_matricula matri
                                        inner join sga_materiaasignada ma on ma.matricula_id=matri.id
                                        inner join sga_materia mate on mate.id=ma.materia_id
                                        inner join sga_asignatura a on a.id=mate.asignatura_id
                                        where matri.id=mat.id order by ma.id  limit 1 offset 2) as asignatura3,
                                        (select a.nombre
                                        from sga_matricula matri
                                        inner join sga_materiaasignada ma on ma.matricula_id=matri.id
                                        inner join sga_materia mate on mate.id=ma.materia_id
                                        inner join sga_asignatura a on a.id=mate.asignatura_id
                                        where matri.id=mat.id order by ma.id  limit 1 offset 3) as asignatura4
                                        from sga_matricula mat
                                        inner join sga_nivel ni on ni.id=mat.nivel_id  
                                        inner join sga_inscripcion i on i.id=mat.inscripcion_id
                                        inner join sga_persona p on p.id=i.persona_id
                                        inner join sga_carrera car on car.id=i.carrera_id
                                        inner join sga_modalidad mo on mo.id=car.modalidad
                                        inner join sga_coordinacion_carrera cc on cc.carrera_id=car.id
                                        where ni.periodo_id=%s  and cc.coordinacion_id=%s
                                        order by p.apellido1, p.apellido2
                                    """ % (ePeriodo.id, eCoordinacion.id)
            cursor.execute(sql)
            results = cursor.fetchall()
            row_num = 4
            for r in results:
                i = 0
                cedula = r[0]
                apellidos = r[1]
                nombres = r[2]
                carrera = r[3]
                modalidad = r[4]
                status = r[5]
                promedio = r[6]
                primeravez = r[7]
                segundavez = r[8]
                asignatura1 = r[9]
                asignatura2 = r[10]
                asignatura3 = r[11]
                asignatura4 = r[12]

                ws.write(row_num, 0, u'%s' % cedula, formatocelda)
                ws.write(row_num, 1, u'%s' % apellidos, formatocelda)
                ws.write(row_num, 2, u'%s' % nombres, formatocelda)
                ws.write(row_num, 3, u'%s' % carrera, formatocelda)
                ws.write(row_num, 4, u'%s' % modalidad, formatocelda)
                ws.write(row_num, 5, u'%s' % status, formatocelda)
                ws.write(row_num, 6, promedio, formatocelda)
                ws.write(row_num, 7, u'%s' % asignatura1 if asignatura1 else '', formatocelda)
                ws.write(row_num, 8, u'%s' % asignatura2 if asignatura2 else '', formatocelda)
                ws.write(row_num, 9, u'%s' % asignatura3 if asignatura3 else '', formatocelda)
                ws.write(row_num, 10, u'%s' % asignatura4 if asignatura4 else '', formatocelda)
                ws.write(row_num, 11, u'%s' % 'PRIMERA' if primeravez else 'SEGUNDA', formatocelda)
                row_num += 1

            workbook.close()
            if notificacion_id > 0:
                eNotificacion = Notificacion.objects.get(pk=notificacion_id)
                eNotificacion.en_proceso = False
                eNotificacion.cuerpo = f'Generación del reporte de resultados finales de nivelación listo'
                eNotificacion.url = "{}reportes/niveles/{}".format(MEDIA_URL, nombre_archivo)
                eNotificacion.save()
            else:
                eNotificacion = Notificacion(cuerpo=f'Generación del reporte de resultados finales de nivelación listo',
                                             titulo=f'Reporte de resultados finales de nivelación',
                                             destinatario=ePersona,
                                             url="{}reportes/niveles/{}".format(MEDIA_URL, nombre_archivo),
                                             prioridad=1,
                                             app_label='SGA',
                                             fecha_hora_visible=datetime.now() + timedelta(days=1),
                                             tipo=2,
                                             en_proceso=False)
                eNotificacion.save(request)
            try:
                send_user_notification(user=usernotify,
                                       payload={"head": "Reporte de resultados finales de nivelación",
                                                "body": 'Generación de reporte de resultados finales de nivelación listo',
                                                "action": "notificacion",
                                                "timestamp": time.mktime(datetime.now().timetuple()),
                                                "url": "{}reportes/niveles/{}".format(MEDIA_URL, nombre_archivo),
                                                "btn_notificaciones": traerNotificaciones(request, data, ePersona),
                                                "mensaje": 'Su reporte ha sido generado con exito'
                                                },
                                       ttl=500)
            except:
                pass
        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)
            if notificacion_id > 0:
                eNotificacion = Notificacion.objects.get(pk=notificacion_id)
                eNotificacion.en_proceso = False
                eNotificacion.error = True
                eNotificacion.titulo = f'Reporte de resultados finales de nivelación falló en la ejecución'
                eNotificacion.cuerpo = textoerror
                eNotificacion.save()
            else:
                eNotificacion = Notificacion(cuerpo=textoerror,
                                             titulo=f'Reporte de resultados finales de nivelación falló en la ejecución',
                                             destinatario=ePersona,
                                             prioridad=1,
                                             app_label='SGA',
                                             fecha_hora_visible=datetime.now() + timedelta(days=1),
                                             tipo=2,
                                             en_proceso=False,
                                             error=True)
                eNotificacion.save(request)
            try:
                send_user_notification(user=usernotify,
                                       payload={"head": "Reporte de resultados finales de nivelación Fallido",
                                                "body": 'Generación del reporte de resultados finales de nivelación a fallado',
                                                "action": "notificacion",
                                                "timestamp": time.mktime(datetime.now().timetuple()),
                                                "btn_notificaciones": traerNotificaciones(request, data, ePersona),
                                                "mensaje": textoerror,
                                                "error": True
                                                },
                                       ttl=500)
            except:
                pass


class reporte_horas_cumplidas_practica_vinculacion_coordinacion(threading.Thread):

    def __init__(self, request, data, notificacion_id):
        self.request = request
        self.data = data
        self.notificacion_id = notificacion_id
        threading.Thread.__init__(self)

    def run(self):
        from sga.models import Coordinacion, Periodo, Persona, Notificacion, Matricula, Inscripcion, \
            PracticasPreprofesionalesInscripcion, ParticipantesMatrices
        from sga.funciones import null_to_numeric
        try:
            directory = os.path.join(MEDIA_ROOT, 'reportes', 'niveles')
            request, data, notificacion_id = self.request, self.data, self.notificacion_id
            usernotify = User.objects.get(pk=request.user.pk)
            ePersona = Persona.objects.get(usuario=usernotify)
            os.makedirs(directory, exist_ok=True)

            try:
                ePeriodo = Periodo.objects.get(pk=int(request.GET.get('idp', '0')))
            except ObjectDoesNotExist:
                raise NameError(u"No se encontro coordinacion académica")
            try:
                eCoordinacion = Coordinacion.objects.get(pk=int(request.GET.get('idc', '0')))
            except ObjectDoesNotExist:
                raise NameError(u"No se encontro coordinacion académica")

            nombre_archivo = "rpt_horas_cumplidas_practica_vinculacion_coordinacion_{}_{}.xlsx".format(eCoordinacion.alias.lower() if eCoordinacion.alias else eCoordinacion.nombre.lower(), random.randint(1, 10000).__str__())
            directory = os.path.join(MEDIA_ROOT, 'reportes', 'niveles', nombre_archivo)
            workbook = xlsxwriter.Workbook(directory, {'constant_memory': True})
            ws = workbook.add_worksheet('Hoja1')
            formatocabeceracolumna = workbook.add_format({
                'bold': 1,
                'border': 1,
                'align': 'center',
                'valign': 'vcenter',
                'bg_color': 'silver',
                'text_wrap': 1,
                'font_size': 10})

            formatocelda = workbook.add_format({
                'border': 1
            })

            formatotitulo = workbook.add_format(
                {'align': 'center', 'valign': 'vcenter', 'bold': 1, 'font_size': 14})

            ws.merge_range(0, 0, 0, 14, 'UNIVERSIDAD ESTATAL DE MILAGRO', formatotitulo)
            ws.merge_range(1, 0, 1, 14, 'DIRECCIÓN DE VINCULACIÓN', formatotitulo)
            ws.merge_range(2, 0, 2, 14, 'HORAS CUMPLIDAS DE PRÁCTICAS LABORALES Y VINCULACIÓN', formatotitulo)
            ws.merge_range(3, 0, 3, 14, eCoordinacion.nombre + " - " + eCoordinacion.alias, formatotitulo)

            columns = [
                (u"MALLA", 20),
                (u"FACULTAD", 40),
                (u"CARRERA", 40),
                (u"NIVEL MALLA", 40),
                (u"MODALIDAD", 20),
                (u"CÉDULA", 20),
                (u"ESTUDIANTE", 15),
                (u"CORREO INSTITUCIONAL", 35),
                (u"CORREO PERSONAL", 35),
                (u"CELULAR", 35),
                (u"CONVENCIONAL", 35),
                (u"CANTÓN", 35),
                (u"HORAS PRÁCTICAS CULMINADAS", 35),
                (u"HORAS PRÁCTICAS HOMOLOGADAS", 35),
                (u"HORAS PRÁCTICAS EN ESTADO APROBADO O PENDIENTE", 35),
                (u"HORAS PRÁCTICAS REQUISITO", 35),
                (u"HORAS VINCULACIÓN ", 35),
                (u"HORAS VINCULACIÓN REQUISITO", 35),
                (u"ASIGNATURAS PENDIENTES", 35)
            ]
            row_num = 4
            for col_num in range(len(columns)):
                ws.write(row_num, col_num, columns[col_num][0], formatocabeceracolumna)
                ws.set_column(col_num, col_num, columns[col_num][1])
            filtro = Q(status=True) & Q(nivel__periodo_id=ePeriodo.id) \
                     & Q(nivel__nivellibrecoordinacion__coordinacion_id=eCoordinacion.id) \
                     & Q(estado_matricula__in=[2, 3]) & Q(inscripcion__inscripcionnivel__status=True) \
                     & Q(inscripcion__inscripcionnivel__nivel__status=True) \
                     & Q(inscripcion__inscripcionnivel__nivel__orden__gte=7)
            eMatriculas = Matricula.objects.filter(filtro).exclude(retiradomatricula=True)

            total = eMatriculas.values("id").count()

            row_num = 5
            for eMatricula in eMatriculas:
                eInscripcion = eMatricula.inscripcion
                nivelmalla = eInscripcion.mi_nivel().nivel

                if nivelmalla.orden >= 7:
                    malla = eInscripcion.carrera.malla()
                    horas_practica_malla = malla.horas_practicas
                    horas_vinculacion_malla = malla.horas_vinculacion

                    facultad = eMatricula.nivel.coordinacion()
                    carrera = eInscripcion.carrera.nombre

                    modalidad = eInscripcion.modalidad
                    persona = eInscripcion.persona

                    cedula = persona.identificacion()
                    nombres = persona.nombre_completo_inverso()
                    correo_inst = persona.emailinst
                    correo_pers = persona.email
                    celular = persona.telefono
                    telefono = persona.telefono_conv
                    canton = persona.canton.nombre

                    practicas = PracticasPreprofesionalesInscripcion.objects.filter(inscripcion=eInscripcion, status=True, retirado=False).distinct().order_by('-fecha_creacion')
                    horas_culminadas = null_to_numeric(practicas.filter(~Q(tiposolicitud=3), culminada=True).aggregate(totalhora=Sum('numerohora'))['totalhora'], 0)
                    horas_culminadas_homo = null_to_numeric(practicas.filter(tiposolicitud=3, culminada=True).aggregate(totalhorahomo=Sum('horahomologacion'))['totalhorahomo'], 0)
                    total_horas_practica_culminada = int(horas_culminadas) + int(horas_culminadas_homo)

                    horas_aprobadas_pend = null_to_numeric(practicas.filter(~Q(tiposolicitud=3), culminada=False, estadosolicitud__in=[2, 4]).aggregate(totalhora=Sum('numerohora'))['totalhora'], 0)
                    horas_aprobadas_pend_homo = null_to_numeric(practicas.filter(tiposolicitud=3, culminada=False, estadosolicitud__in=[2, 4]).aggregate(totalhorahomo=Sum('horahomologacion'))['totalhorahomo'], 0)
                    total_horas_aprobada_pend = int(horas_aprobadas_pend) + int(horas_aprobadas_pend_homo)

                    total_horas_vinculacion = null_to_numeric(ParticipantesMatrices.objects.values('horas').filter(matrizevidencia_id=2, status=True, proyecto__status=True, proyecto__tipo=1, inscripcion__persona=persona).aggregate(totalhoravincu=Sum('horas'))['totalhoravincu'], 0)

                    materias_pendientes = eInscripcion.cantidad_materias_pendiente_aprobar()

                    ws.write(row_num, 0, u'%s' % malla, formatocelda)
                    ws.write(row_num, 1, u'%s' % facultad, formatocelda)
                    ws.write(row_num, 2, u'%s' % carrera, formatocelda)
                    ws.write(row_num, 3, u'%s' % nivelmalla, formatocelda)
                    ws.write(row_num, 4, u'%s' % modalidad, formatocelda)
                    ws.write(row_num, 5, u'%s' % cedula, formatocelda)
                    ws.write(row_num, 6, u'%s' % nombres, formatocelda)
                    ws.write(row_num, 7, u'%s' % correo_inst, formatocelda)
                    ws.write(row_num, 8, u'%s' % correo_pers, formatocelda)
                    ws.write(row_num, 9, u'%s' % celular, formatocelda)
                    ws.write(row_num, 10, u'%s' % telefono, formatocelda)
                    ws.write(row_num, 11, u'%s' % canton, formatocelda)
                    if eInscripcion.exonerado_practias():
                        ws.write(row_num, 12, 'EXONERADO', formatocelda)
                    else:
                        ws.write(row_num, 12, u'%s' % total_horas_practica_culminada, formatocelda)
                    if eInscripcion.exonerado_practias():
                        ws.write(row_num, 13, 'EXONERADO', formatocelda)
                    else:
                        ws.write(row_num, 13, u'%s' % horas_culminadas_homo, formatocelda)
                    if eInscripcion.exonerado_practias():
                        ws.write(row_num, 14, 'EXONERADO', formatocelda)
                    else:
                        ws.write(row_num, 14, u'%s' % total_horas_aprobada_pend, formatocelda)
                    if eInscripcion.exonerado_practias():
                        ws.write(row_num, 15, 'EXONERADO', formatocelda)
                    else:
                        ws.write(row_num, 15, u'%s' % horas_practica_malla, formatocelda)
                    if eInscripcion.exonerado_practias():
                        ws.write(row_num, 16, 'EXONERADO', formatocelda)
                    else:
                        ws.write(row_num, 16, u'%s' % total_horas_vinculacion, formatocelda)
                    ws.write(row_num, 17, u'%s' % horas_vinculacion_malla, formatocelda)
                    ws.write(row_num, 18, u'%s' % materias_pendientes, formatocelda)

                    row_num += 1
            workbook.close()
            if notificacion_id > 0:
                eNotificacion = Notificacion.objects.get(pk=notificacion_id)
                eNotificacion.en_proceso = False
                eNotificacion.cuerpo = f'Generación del reporte de horas cumplidas prácticas y vinculación {eCoordinacion.nombre.lower()} listo'
                eNotificacion.url = "{}reportes/niveles/{}".format(MEDIA_URL, nombre_archivo)
                eNotificacion.save()
            else:
                eNotificacion = Notificacion(cuerpo=f'Generación del reporte de horas cumplidas prácticas y vinculación {eCoordinacion.nombre.lower()} listo',
                                             titulo=f'Reporte de horas cumplidas prácticas y vinculación {eCoordinacion.alias if eCoordinacion.alias else eCoordinacion.nombre}',
                                             destinatario=ePersona,
                                             url="{}reportes/niveles/{}".format(MEDIA_URL, nombre_archivo),
                                             prioridad=1,
                                             app_label='SGA',
                                             fecha_hora_visible=datetime.now() + timedelta(days=1),
                                             tipo=2,
                                             en_proceso=False)
                eNotificacion.save(request)
            try:
                send_user_notification(user=usernotify,
                                       payload={"head": "Reporte de horas cumplidas prácticas y vinculación",
                                                "body": 'Generación de reporte de horas cumplidas prácticas y vinculación listo',
                                                "action": "notificacion",
                                                "timestamp": time.mktime(datetime.now().timetuple()),
                                                "url": "{}reportes/niveles/{}".format(MEDIA_URL, nombre_archivo),
                                                "btn_notificaciones": traerNotificaciones(request, data, ePersona),
                                                "mensaje": 'Su reporte ha sido generado con exito'
                                                },
                                       ttl=500)
            except:
                pass
        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)
            if notificacion_id > 0:
                eNotificacion = Notificacion.objects.get(pk=notificacion_id)
                eNotificacion.en_proceso = False
                eNotificacion.error = True
                eNotificacion.titulo = f'Reporte de horas cumplidas prácticas y vinculación {eCoordinacion.nombre.lower()} falló en la ejecución'
                eNotificacion.cuerpo = textoerror
                eNotificacion.save()
            else:
                eNotificacion = Notificacion(cuerpo=textoerror,
                                             titulo=f'Reporte de horas cumplidas prácticas y vinculación {eCoordinacion.nombre.lower()} falló en la ejecución',
                                             destinatario=ePersona,
                                             prioridad=1,
                                             app_label='SGA',
                                             fecha_hora_visible=datetime.now() + timedelta(days=1),
                                             tipo=2,
                                             en_proceso=False,
                                             error=True)
                eNotificacion.save(request)
            try:
                send_user_notification(user=usernotify,
                                       payload={"head": "Reporte de horas cumplidas prácticas y vinculación Fallido",
                                                "body": 'Generación del reporte de horas cumplidas prácticas y vinculación a fallado',
                                                "action": "notificacion",
                                                "timestamp": time.mktime(datetime.now().timetuple()),
                                                "btn_notificaciones": traerNotificaciones(request, data, ePersona),
                                                "mensaje": textoerror,
                                                "error": True
                                                },
                                       ttl=500)
            except:
                pass


class reporte_horas_cumplidas_practica_vinculacion_niveles_coordinacion(threading.Thread):

    def __init__(self, request, data, notificacion_id):
        self.request = request
        self.data = data
        self.notificacion_id = notificacion_id
        threading.Thread.__init__(self)

    def run(self):
        from sga.models import Coordinacion, Periodo, Persona, Notificacion, Matricula, Inscripcion, \
            PracticasPreprofesionalesInscripcion, ParticipantesMatrices
        from sga.funciones import null_to_numeric
        try:
            directory = os.path.join(MEDIA_ROOT, 'reportes', 'niveles')
            request, data, notificacion_id = self.request, self.data, self.notificacion_id
            usernotify = User.objects.get(pk=request.user.pk)
            ePersona = Persona.objects.get(usuario=usernotify)
            os.makedirs(directory, exist_ok=True)

            try:
                ePeriodo = Periodo.objects.get(pk=int(request.GET.get('idp', '0')))
            except ObjectDoesNotExist:
                raise NameError(u"No se encontro coordinacion académica")
            try:
                eCoordinacion = Coordinacion.objects.get(pk=int(request.GET.get('idc', '0')))
            except ObjectDoesNotExist:
                raise NameError(u"No se encontro coordinacion académica")

            nombre_archivo = "rpt_horas_cumplidas_practica_vinculacion_todos_niveles_coordinacion_{}_{}.xlsx".format(eCoordinacion.alias.lower() if eCoordinacion.alias else eCoordinacion.nombre.lower(), random.randint(1, 10000).__str__())
            directory = os.path.join(MEDIA_ROOT, 'reportes', 'niveles', nombre_archivo)
            workbook = xlsxwriter.Workbook(directory, {'constant_memory': True})
            ws = workbook.add_worksheet('Hoja1')
            formatocabeceracolumna = workbook.add_format({
                'bold': 1,
                'border': 1,
                'align': 'center',
                'valign': 'vcenter',
                'bg_color': 'silver',
                'text_wrap': 1,
                'font_size': 10})

            formatocelda = workbook.add_format({
                'border': 1
            })

            formatotitulo = workbook.add_format(
                {'align': 'center', 'valign': 'vcenter', 'bold': 1, 'font_size': 14})

            ws.merge_range(0, 0, 0, 14, 'UNIVERSIDAD ESTATAL DE MILAGRO', formatotitulo)
            ws.merge_range(1, 0, 1, 14, 'DIRECCIÓN DE VINCULACIÓN', formatotitulo)
            ws.merge_range(2, 0, 2, 14, 'HORAS CUMPLIDAS DE PRÁCTICAS LABORALES Y VINCULACIÓN', formatotitulo)
            ws.merge_range(3, 0, 3, 14, eCoordinacion.nombre + " - " + eCoordinacion.alias, formatotitulo)

            columns = [
                (u"MALLA", 20),
                (u"FACULTAD", 40),
                (u"CARRERA", 40),
                (u"NIVEL MALLA", 40),
                (u"MODALIDAD", 20),
                (u"CÉDULA", 20),
                (u"ESTUDIANTE", 15),
                (u"CORREO INSTITUCIONAL", 35),
                (u"CORREO PERSONAL", 35),
                (u"CELULAR", 35),
                (u"CONVENCIONAL", 35),
                (u"PAIS", 35),
                (u"PROVINCIA", 35),
                (u"CANTÓN", 35),
                (u"HORAS PRÁCTICAS CULMINADAS", 35),
                (u"HORAS PRÁCTICAS HOMOLOGADAS", 35),
                (u"HORAS PRÁCTICAS EN ESTADO APROBADO O PENDIENTE", 35),
                (u"HORAS PRÁCTICAS REQUISITO", 35),
                (u"HORAS VINCULACIÓN ", 35),
                (u"HORAS VINCULACIÓN REQUISITO", 35),
                (u"ASIGNATURAS PENDIENTES", 35)
            ]
            row_num = 4
            for col_num in range(len(columns)):
                ws.write(row_num, col_num, columns[col_num][0], formatocabeceracolumna)
                ws.set_column(col_num, col_num, columns[col_num][1])
            matriculas = Matricula.objects.filter(nivel__periodo_id=ePeriodo.id, inscripcion__carrera__coordinacion__id=eCoordinacion.id, status=True, retiradomatricula=False).exclude(inscripcion__coordinacion__id__in=[9, 7, 10])
            total = matriculas.count()
            row_num = 5
            for matricula in matriculas:
                inscripcion = matricula.inscripcion
                nivelmalla = inscripcion.mi_nivel().nivel
                nivemallamatricula = matricula.nivelmalla
                malla = inscripcion.carrera.malla()
                horas_practica_malla = malla.horas_practicas
                horas_vinculacion_malla = malla.horas_vinculacion
                facultad = matricula.nivel.coordinacion()
                carrera = inscripcion.carrera.nombre
                modalidad = inscripcion.modalidad
                persona = inscripcion.persona
                cedula = persona.identificacion()
                nombres = persona.nombre_completo_inverso()
                correo_inst = persona.emailinst
                correo_pers = persona.email
                celular = persona.telefono
                telefono = persona.telefono_conv
                pais = '' if not persona.pais else persona.pais.nombre
                provincia = '' if not persona.provincia else persona.provincia.nombre
                canton = '' if not persona.canton else persona.canton.nombre

                practicas = PracticasPreprofesionalesInscripcion.objects.filter(inscripcion=inscripcion, status=True, retirado=False).distinct().order_by('-fecha_creacion')
                horas_culminadas = null_to_numeric(practicas.filter(~Q(tiposolicitud=3), culminada=True).aggregate(totalhora=Sum('numerohora'))['totalhora'], 0)
                horas_culminadas_homo = null_to_numeric(practicas.filter(tiposolicitud=3, culminada=True).aggregate(totalhorahomo=Sum('horahomologacion'))['totalhorahomo'], 0)
                total_horas_practica_culminada = int(horas_culminadas) + int(horas_culminadas_homo)
                horas_aprobadas_pend = null_to_numeric(practicas.filter(~Q(tiposolicitud=3), culminada=False, estadosolicitud__in=[2, 4]).aggregate(totalhora=Sum('numerohora'))['totalhora'], 0)
                horas_aprobadas_pend_homo = null_to_numeric(practicas.filter(tiposolicitud=3, culminada=False, estadosolicitud__in=[2, 4]).aggregate(totalhorahomo=Sum('horahomologacion'))['totalhorahomo'], 0)
                total_horas_aprobada_pend = int(horas_aprobadas_pend) + int(horas_aprobadas_pend_homo)
                total_horas_vinculacion = null_to_numeric(ParticipantesMatrices.objects.values('horas').filter(matrizevidencia_id=2, status=True, inscripcion=inscripcion).aggregate(totalhoravincu=Sum('horas'))['totalhoravincu'], 0)
                materias_pendientes = inscripcion.cantidad_materias_pendiente_aprobar()
                ws.write(row_num, 0, u'%s' % malla, formatocelda)
                ws.write(row_num, 1, u'%s' % facultad, formatocelda)
                ws.write(row_num, 2, u'%s' % carrera, formatocelda)
                ws.write(row_num, 3, u'%s' % nivelmalla, formatocelda)
                ws.write(row_num, 4, u'%s' % modalidad, formatocelda)
                ws.write(row_num, 5, u'%s' % cedula, formatocelda)
                ws.write(row_num, 6, u'%s' % nombres, formatocelda)
                ws.write(row_num, 7, u'%s' % correo_inst, formatocelda)
                ws.write(row_num, 8, u'%s' % correo_pers, formatocelda)
                ws.write(row_num, 9, u'%s' % celular, formatocelda)
                ws.write(row_num, 10, u'%s' % telefono, formatocelda)
                ws.write(row_num, 11, u'%s' % pais, formatocelda)
                ws.write(row_num, 12, u'%s' % provincia, formatocelda)
                ws.write(row_num, 13, u'%s' % canton, formatocelda)
                if inscripcion.exonerado_practias():
                    ws.write(row_num, 14, 'EXONERADO', formatocelda)
                else:
                    ws.write(row_num, 14, u'%s' % total_horas_practica_culminada, formatocelda)
                if inscripcion.exonerado_practias():
                    ws.write(row_num, 15, 'EXONERADO', formatocelda)
                else:
                    ws.write(row_num, 15, u'%s' % horas_culminadas_homo, formatocelda)
                if inscripcion.exonerado_practias():
                    ws.write(row_num, 16, 'EXONERADO', formatocelda)
                else:
                    ws.write(row_num, 16, u'%s' % total_horas_aprobada_pend, formatocelda)
                if inscripcion.exonerado_practias():
                    ws.write(row_num, 17, 'EXONERADO', formatocelda)
                else:
                    ws.write(row_num, 17, u'%s' % horas_practica_malla, formatocelda)
                if inscripcion.exonerado_practias():
                    ws.write(row_num, 18, 'EXONERADO', formatocelda)
                else:
                    ws.write(row_num, 18, u'%s' % total_horas_vinculacion, formatocelda)
                ws.write(row_num, 19, u'%s' % horas_vinculacion_malla, formatocelda)
                ws.write(row_num, 20, u'%s' % materias_pendientes, formatocelda)
                ws.write(row_num, 21, u'%s' % nivemallamatricula, formatocelda)

                row_num += 1
            workbook.close()
            if notificacion_id > 0:
                eNotificacion = Notificacion.objects.get(pk=notificacion_id)
                eNotificacion.en_proceso = False
                eNotificacion.cuerpo = f'Generación del reporte de horas cumplidas prácticas y vinculación todos los niveles {eCoordinacion.nombre.lower()} listo'
                eNotificacion.url = "{}reportes/niveles/{}".format(MEDIA_URL, nombre_archivo)
                eNotificacion.save()
            else:
                eNotificacion = Notificacion(cuerpo=f'Generación del reporte de horas cumplidas prácticas y vinculación todos los niveles {eCoordinacion.nombre.lower()} listo',
                                             titulo=f'Reporte de horas cumplidas prácticas y vinculación todos los niveles {eCoordinacion.alias if eCoordinacion.alias else eCoordinacion.nombre}',
                                             destinatario=ePersona,
                                             url="{}reportes/niveles/{}".format(MEDIA_URL, nombre_archivo),
                                             prioridad=1,
                                             app_label='SGA',
                                             fecha_hora_visible=datetime.now() + timedelta(days=1),
                                             tipo=2,
                                             en_proceso=False)
                eNotificacion.save(request)
            try:
                send_user_notification(user=usernotify,
                                       payload={"head": "Reporte de horas cumplidas prácticas y vinculación todos los niveles",
                                                "body": 'Generación de reporte de horas cumplidas prácticas y vinculación todos los niveles listo',
                                                "action": "notificacion",
                                                "timestamp": time.mktime(datetime.now().timetuple()),
                                                "url": "{}reportes/niveles/{}".format(MEDIA_URL, nombre_archivo),
                                                "btn_notificaciones": traerNotificaciones(request, data, ePersona),
                                                "mensaje": 'Su reporte ha sido generado con exito'
                                                },
                                       ttl=500)
            except:
                pass
        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)
            if notificacion_id > 0:
                eNotificacion = Notificacion.objects.get(pk=notificacion_id)
                eNotificacion.en_proceso = False
                eNotificacion.error = True
                eNotificacion.titulo = f'Reporte de horas cumplidas prácticas y vinculación todos los niveles {eCoordinacion.nombre.lower()} falló en la ejecución'
                eNotificacion.cuerpo = textoerror
                eNotificacion.save()
            else:
                eNotificacion = Notificacion(cuerpo=textoerror,
                                             titulo=f'Reporte de horas cumplidas prácticas y vinculación todos los niveles {eCoordinacion.nombre.lower()} falló en la ejecución',
                                             destinatario=ePersona,
                                             prioridad=1,
                                             app_label='SGA',
                                             fecha_hora_visible=datetime.now() + timedelta(days=1),
                                             tipo=2,
                                             en_proceso=False,
                                             error=True)
                eNotificacion.save(request)
            try:
                send_user_notification(user=usernotify,
                                       payload={"head": "Reporte de horas cumplidas prácticas y vinculación todos los niveles Fallido",
                                                "body": 'Generación del reporte de horas cumplidas prácticas y vinculación todos los niveles a fallado',
                                                "action": "notificacion",
                                                "timestamp": time.mktime(datetime.now().timetuple()),
                                                "btn_notificaciones": traerNotificaciones(request, data, ePersona),
                                                "mensaje": textoerror,
                                                "error": True
                                                },
                                       ttl=500)
            except:
                pass


class reporte_estudiantes_practicas_pendientes_matriculados_coordinacion(threading.Thread):

    def __init__(self, request, data, notificacion_id):
        self.request = request
        self.data = data
        self.notificacion_id = notificacion_id
        threading.Thread.__init__(self)

    def run(self):
        from sga.models import Coordinacion, Periodo, Persona, Notificacion, Matricula, Inscripcion, \
            PracticasPreprofesionalesInscripcion, ParticipantesMatrices
        from sga.funciones import null_to_numeric
        try:
            directory = os.path.join(MEDIA_ROOT, 'reportes', 'niveles')
            request, data, notificacion_id = self.request, self.data, self.notificacion_id
            usernotify = User.objects.get(pk=request.user.pk)
            ePersona = Persona.objects.get(usuario=usernotify)
            os.makedirs(directory, exist_ok=True)

            try:
                ePeriodo = Periodo.objects.get(pk=int(request.GET.get('idp', '0')))
            except ObjectDoesNotExist:
                raise NameError(u"No se encontro coordinacion académica")
            try:
                eCoordinacion = Coordinacion.objects.get(pk=int(request.GET.get('idc', '0')))
            except ObjectDoesNotExist:
                raise NameError(u"No se encontro coordinacion académica")

            nombre_archivo = "rpt_estudiantes_practicas_pendientes_matriculados_coordinacion_{}_{}.xlsx".format(eCoordinacion.alias.lower() if eCoordinacion.alias else eCoordinacion.nombre.lower(), random.randint(1, 10000).__str__())
            directory = os.path.join(MEDIA_ROOT, 'reportes', 'niveles', nombre_archivo)
            workbook = xlsxwriter.Workbook(directory, {'constant_memory': True})
            ws = workbook.add_worksheet('Hoja1')
            cn = connections['sga_select'].cursor()
            sql = """   SELECT MIN(nm.orden) AS nivelminimo
                        FROM sga_itinerariosmalla AS i
                            INNER JOIN sga_nivelmalla AS nm ON i.nivel_id=nm.id
                            INNER JOIN sga_malla AS ma ON i.malla_id=ma.id
                            INNER JOIN sga_carrera AS ca ON ma.carrera_id=ca.id
                            INNER JOIN sga_coordinacion_carrera AS cc ON cc.carrera_id=ca.id
                            INNER JOIN sga_coordinacion AS fa ON cc.coordinacion_id=fa.id
                        WHERE i.status=TRUE AND fa.id=%s """ % eCoordinacion.id

            cn.execute(sql)
            registro = cn.fetchall()
            nivelminimoiti = registro[0][0] if registro[0][0] else 6
            cn.close()

            formatocabeceracolumna = workbook.add_format({
                'bold': 1,
                'border': 1,
                'align': 'center',
                'valign': 'vcenter',
                'bg_color': 'silver',
                'text_wrap': 1,
                'font_size': 10})

            formatocelda = workbook.add_format({
                'border': 1
            })

            formatotitulo = workbook.add_format(
                {'align': 'center', 'valign': 'vcenter', 'bold': 1, 'font_size': 14})

            ws.merge_range(0, 0, 0, 12, 'UNIVERSIDAD ESTATAL DE MILAGRO', formatotitulo)
            ws.merge_range(1, 0, 1, 12, 'DIRECCIÓN DE VINCULACIÓN', formatotitulo)
            ws.merge_range(2, 0, 2, 12, 'LISTADO DE ESTUDIANTES MATRICULADOS CON PRÁCTICAS PENDIENTES', formatotitulo)
            ws.merge_range(3, 0, 3, 12, eCoordinacion.nombre + " - " + eCoordinacion.alias, formatotitulo)

            columns = [
                (u"MALLA", 20),
                (u"FACULTAD", 40),
                (u"CARRERA", 40),
                (u"NIVEL MALLA", 20),
                (u"MODALIDAD", 20),
                (u"CÉDULA", 20),
                (u"ESTUDIANTE", 40),
                (u"CORREO INSTITUCIONAL", 35),
                (u"CORREO PERSONAL", 35),
                (u"CELULAR", 20),
                (u"CONVENCIONAL", 20),
                (u"CANTÓN", 20),
                (u"HORAS PRÁCTICAS REQUISITO", 20),
                (u"HORAS PRÁCTICAS CULMINADAS", 20),
                (u"HORAS PRÁCTICAS PENDIENTES", 20),
                (u"ASIGNATURAS PENDIENTES", 20),
                (u"OBSERVACIÓN", 50)
            ]
            row_num = 4
            for col_num in range(len(columns)):
                ws.write(row_num, col_num, columns[col_num][0], formatocabeceracolumna)
                ws.set_column(col_num, col_num, columns[col_num][1])

            matriculas = Matricula.objects.filter(nivel__periodo_id=ePeriodo.id, nivel__nivellibrecoordinacion__coordinacion_id=eCoordinacion.id, estado_matricula__in=[2, 3], status=True, inscripcion__inscripcionnivel__status=True, inscripcion__inscripcionnivel__nivel__status=True, inscripcion__inscripcionnivel__nivel__orden__gte=nivelminimoiti).exclude(retiradomatricula=True).order_by('inscripcion__persona__apellido1', 'inscripcion__persona__apellido2', 'inscripcion__persona__nombres')
            row_num = 5
            c = 1
            aptos = 0
            matriculado = True
            for matricula in matriculas:
                inscripcion = matricula.inscripcion
                nivelmalla = inscripcion.mi_nivel().nivel

                # Itinerario de la malla
                itinerario = inscripcion.inscripcionmalla_set.filter(status=True).first().malla.itinerariosmalla_set.filter(status=True).order_by('nivel__orden')
                if itinerario:
                    nivelminimoitinerario = itinerario[0].nivel.orden
                    verificar_apto = nivelmalla.orden >= nivelminimoitinerario
                else:
                    verificar_apto = nivelmalla.orden > 5

                if verificar_apto:
                    # print("Revisando ", c)
                    c += 1
                    apto = inscripcion.apto_para_realizar_practicas(matriculado)['apto']
                    # print(apto)
                    if apto:
                        aptos += 1
                        # print("Aptos para prácticas:", aptos)
                        # //malla = inscripcion.carrera.malla()
                        malla = inscripcion.mi_malla()
                        horas_practica_malla = malla.horas_practicas
                        facultad = matricula.nivel.coordinacion()
                        carrera = inscripcion.carrera.nombre
                        modalidad = inscripcion.modalidad
                        persona = inscripcion.persona
                        cedula = persona.identificacion()
                        nombres = persona.nombre_completo_inverso()
                        correo_inst = persona.emailinst
                        correo_pers = persona.email
                        celular = persona.telefono
                        telefono = persona.telefono_conv
                        canton = persona.canton.nombre if persona.canton else ''
                        practicas = PracticasPreprofesionalesInscripcion.objects.filter(inscripcion=inscripcion, status=True, retirado=False).distinct().order_by('-fecha_creacion')
                        horas_culminadas = null_to_numeric(practicas.filter(~Q(tiposolicitud=3), culminada=True).aggregate(totalhora=Sum('numerohora'))['totalhora'], 0)
                        horas_culminadas_homo = null_to_numeric(practicas.filter(tiposolicitud=3, culminada=True).aggregate(totalhorahomo=Sum('horahomologacion'))['totalhorahomo'], 0)
                        total_horas_practica_culminada = int(horas_culminadas) + int(horas_culminadas_homo)
                        horas_pendientes = 0
                        if horas_practica_malla > 0:
                            horas_pendientes = horas_practica_malla - total_horas_practica_culminada
                        materias_pendientes = inscripcion.cantidad_materias_pendiente_aprobar()
                        ws.write(row_num, 0, u'%s' % malla, formatocelda)
                        ws.write(row_num, 1, u'%s' % facultad, formatocelda)
                        ws.write(row_num, 2, u'%s' % carrera, formatocelda)
                        ws.write(row_num, 3, u'%s' % nivelmalla, formatocelda)
                        ws.write(row_num, 4, u'%s' % modalidad, formatocelda)
                        ws.write(row_num, 5, u'%s' % cedula, formatocelda)
                        ws.write(row_num, 6, u'%s' % nombres, formatocelda)
                        ws.write(row_num, 7, u'%s' % correo_inst, formatocelda)
                        ws.write(row_num, 8, u'%s' % correo_pers, formatocelda)
                        ws.write(row_num, 9, u'%s' % celular, formatocelda)
                        ws.write(row_num, 10, u'%s' % telefono, formatocelda)
                        ws.write(row_num, 11, u'%s' % canton, formatocelda)
                        ws.write(row_num, 12, u'%s' % horas_practica_malla, formatocelda)
                        ws.write(row_num, 13, u'%s' % total_horas_practica_culminada, formatocelda)
                        ws.write(row_num, 14, u'%s' % horas_pendientes, formatocelda)
                        ws.write(row_num, 15, u'%s' % materias_pendientes, formatocelda)
                        ws.write(row_num, 16, u'%s' % "HORAS PRÁCTICAS NO CONFIGURADAS EN LA MALLA" if horas_practica_malla == 0 else "", formatocelda)

                        row_num += 1
                    else:
                        print("No apto ", inscripcion.persona.cedula)
                else:
                    print("No califica ni para revision...", inscripcion.persona.cedula)
            workbook.close()
            if notificacion_id > 0:
                eNotificacion = Notificacion.objects.get(pk=notificacion_id)
                eNotificacion.en_proceso = False
                eNotificacion.cuerpo = f'Generación del reporte de estudiantes con prácticas pendientes (Matriculados) {eCoordinacion.nombre.lower()} listo'
                eNotificacion.url = "{}reportes/niveles/{}".format(MEDIA_URL, nombre_archivo)
                eNotificacion.save()
            else:
                eNotificacion = Notificacion(cuerpo=f'Generación del reporte de estudiantes con prácticas pendientes (Matriculados) {eCoordinacion.nombre.lower()} listo',
                                             titulo=f'Reporte de estudiantes con prácticas pendientes (Matriculados) {eCoordinacion.alias if eCoordinacion.alias else eCoordinacion.nombre}',
                                             destinatario=ePersona,
                                             url="{}reportes/niveles/{}".format(MEDIA_URL, nombre_archivo),
                                             prioridad=1,
                                             app_label='SGA',
                                             fecha_hora_visible=datetime.now() + timedelta(days=1),
                                             tipo=2,
                                             en_proceso=False)
                eNotificacion.save(request)
            try:
                send_user_notification(user=usernotify,
                                       payload={"head": "Reporte de estudiantes con prácticas pendientes (Matriculados)",
                                                "body": 'Generación de reporte de estudiantes con prácticas pendientes (Matriculados) listo',
                                                "action": "notificacion",
                                                "timestamp": time.mktime(datetime.now().timetuple()),
                                                "url": "{}reportes/niveles/{}".format(MEDIA_URL, nombre_archivo),
                                                "btn_notificaciones": traerNotificaciones(request, data, ePersona),
                                                "mensaje": 'Su reporte ha sido generado con exito'
                                                },
                                       ttl=500)
            except:
                pass
        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)
            if notificacion_id > 0:
                eNotificacion = Notificacion.objects.get(pk=notificacion_id)
                eNotificacion.en_proceso = False
                eNotificacion.error = True
                eNotificacion.titulo = f'Reporte de estudiantes con prácticas pendientes (Matriculados) {eCoordinacion.nombre.lower()} falló en la ejecución'
                eNotificacion.cuerpo = textoerror
                eNotificacion.save()
            else:
                eNotificacion = Notificacion(cuerpo=textoerror,
                                             titulo=f'Reporte de estudiantes con prácticas pendientes (Matriculados) {eCoordinacion.nombre.lower()} falló en la ejecución',
                                             destinatario=ePersona,
                                             prioridad=1,
                                             app_label='SGA',
                                             fecha_hora_visible=datetime.now() + timedelta(days=1),
                                             tipo=2,
                                             en_proceso=False,
                                             error=True)
                eNotificacion.save(request)
            try:
                send_user_notification(user=usernotify,
                                       payload={"head": "Reporte de estudiantes con prácticas pendientes (Matriculados) Fallido",
                                                "body": 'Generación del reporte estudiantes con prácticas pendientes (Matriculados) a fallado',
                                                "action": "notificacion",
                                                "timestamp": time.mktime(datetime.now().timetuple()),
                                                "btn_notificaciones": traerNotificaciones(request, data, ePersona),
                                                "mensaje": textoerror,
                                                "error": True
                                                },
                                       ttl=500)
            except:
                pass


class reporte_estudiantes_practicas_pendientes_no_matriculados_coordinacion(threading.Thread):

    def __init__(self, request, data, notificacion_id):
        self.request = request
        self.data = data
        self.notificacion_id = notificacion_id
        threading.Thread.__init__(self)

    def run(self):
        from sga.models import Coordinacion, Periodo, Persona, Notificacion, Matricula, Inscripcion, \
            PracticasPreprofesionalesInscripcion, ParticipantesMatrices
        from sga.funciones import null_to_numeric
        try:
            directory = os.path.join(MEDIA_ROOT, 'reportes', 'niveles')
            request, data, notificacion_id = self.request, self.data, self.notificacion_id
            usernotify = User.objects.get(pk=request.user.pk)
            ePersona = Persona.objects.get(usuario=usernotify)
            os.makedirs(directory, exist_ok=True)

            try:
                ePeriodo = Periodo.objects.get(pk=int(request.GET.get('idp', '0')))
            except ObjectDoesNotExist:
                raise NameError(u"No se encontro coordinacion académica")
            try:
                eCoordinacion = Coordinacion.objects.get(pk=int(request.GET.get('idc', '0')))
            except ObjectDoesNotExist:
                raise NameError(u"No se encontro coordinacion académica")

            nombre_archivo = "rpt_estudiantes_practicas_pendientes_no_matriculados_coordinacion_{}_{}.xlsx".format(eCoordinacion.alias.lower() if eCoordinacion.alias else eCoordinacion.nombre.lower(), random.randint(1, 10000).__str__())
            directory = os.path.join(MEDIA_ROOT, 'reportes', 'niveles', nombre_archivo)
            workbook = xlsxwriter.Workbook(directory, {'constant_memory': True})
            ws = workbook.add_worksheet('Hoja1')
            formatocabeceracolumna = workbook.add_format({
                'bold': 1,
                'border': 1,
                'align': 'center',
                'valign': 'vcenter',
                'bg_color': 'silver',
                'text_wrap': 1,
                'font_size': 10})

            formatocelda = workbook.add_format({
                'border': 1
            })

            formatotitulo = workbook.add_format(
                {'align': 'center', 'valign': 'vcenter', 'bold': 1, 'font_size': 14})

            ws.merge_range(0, 0, 0, 12, 'UNIVERSIDAD ESTATAL DE MILAGRO', formatotitulo)
            ws.merge_range(1, 0, 1, 12, 'DIRECCIÓN DE VINCULACIÓN', formatotitulo)
            ws.merge_range(2, 0, 2, 12, 'LISTADO DE ESTUDIANTES NO MATRICULADOS CON PRÁCTICAS PENDIENTES', formatotitulo)
            ws.merge_range(3, 0, 3, 12, eCoordinacion.nombre + " - " + eCoordinacion.alias, formatotitulo)

            columns = [
                (u"MALLA", 20),
                (u"FACULTAD", 40),
                (u"CARRERA", 40),
                (u"NIVEL MALLA", 20),
                (u"MODALIDAD", 20),
                (u"CÉDULA", 20),
                (u"ESTUDIANTE", 40),
                (u"CORREO INSTITUCIONAL", 35),
                (u"CORREO PERSONAL", 35),
                (u"CELULAR", 20),
                (u"CONVENCIONAL", 20),
                (u"CANTÓN", 20),
                (u"HORAS PRÁCTICAS REQUISITO", 20),
                (u"HORAS PRÁCTICAS CULMINADAS", 20),
                (u"HORAS PRÁCTICAS PENDIENTES", 20),
                (u"ASIGNATURAS PENDIENTES", 20),
                (u"OBSERVACIÓN", 50)
            ]
            row_num = 4
            for col_num in range(len(columns)):
                ws.write(row_num, col_num, columns[col_num][0], formatocabeceracolumna)
                ws.set_column(col_num, col_num, columns[col_num][1])

            cn = connections['sga_select'].cursor()
            sql = """   SELECT DISTINCT ins.id,pe.apellido1,pe.apellido2,pe.nombres,f.nombre
                        FROM sga_inscripcion AS ins
                            INNER JOIN sga_matricula AS matria ON matria.inscripcion_id=ins.id
                            INNER JOIN sga_carrera AS ca ON ca.id=ins.carrera_id
                            INNER JOIN sga_perfilusuario AS pu ON pu.inscripcion_id=ins.id
                            INNER JOIN sga_persona AS pe ON ins.persona_id=pe.id
                            INNER JOIN sga_inscripcionmalla AS insmall ON insmall.inscripcion_id=ins.id
                            INNER JOIN sga_malla AS mall ON mall.id=insmall.malla_id
                            INNER JOIN sga_inscripcionnivel AS insniv ON insniv.inscripcion_id=ins.id
                            INNER JOIN sga_nivelmalla AS nm ON insniv.nivel_id=nm.id
                            INNER JOIN sga_coordinacion_carrera AS cc ON cc.carrera_id=ca.id
                            INNER JOIN sga_coordinacion AS f ON cc.coordinacion_id=f.id
                        WHERE ins."status"=TRUE AND pe."status"=true AND pu.visible=true AND cc.coordinacion_id=%s
                                AND (EXTRACT(YEAR FROM NOW()) - extract(year from matria.fecha))<=5
                                AND ins.id NOT IN (SELECT retiro.inscripcion_id FROM sga_retirocarrera as retiro Where retiro.status=TRUE)
                                AND ins.id NOT IN (SELECT gra.inscripcion_id FROM sga_graduado AS gra WHERE gra.status=TRUE and gra.estadograduado=true)
                                AND ins.id NOT IN (SELECT matri.inscripcion_id 
                                                    FROM sga_matricula AS matri
                                                        INNER join sga_nivel AS ni ON matri.nivel_id=ni.id
                                                    WHERE matri.estado_matricula in (2,3)
                                                            AND matri.retiradomatricula=FALSE
                                                            and ni.periodo_id=%s)
                                ORDER BY pe.apellido1,pe.apellido2,pe.nombres 	                
                                """ % (eCoordinacion.id, ePeriodo.id)

            cn.execute(sql)
            results = cn.fetchall()

            row_num = 5
            c = 1
            aptos = 0
            matriculado = False

            for r in results:
                idins = r[0]
                facultad = r[4]
                inscripcion = Inscripcion.objects.get(pk=idins)
                nivelmalla = inscripcion.mi_nivel().nivel

                # Itinerario de la malla
                itinerario = inscripcion.inscripcionmalla_set.filter(status=True).first().malla.itinerariosmalla_set.filter(status=True).order_by('nivel__orden')
                if itinerario:
                    nivelminimoitinerario = itinerario[0].nivel.orden
                    verificar_apto = nivelmalla.orden >= nivelminimoitinerario
                else:
                    verificar_apto = nivelmalla.orden > 5

                if verificar_apto:
                    # print("Revisando ", c)
                    c += 1
                    apto = inscripcion.apto_para_realizar_practicas(matriculado)['apto']
                    # print(apto)
                    if apto:
                        aptos += 1
                        # print("Aptos para prácticas:", aptos)
                        # malla = inscripcion.carrera.malla()
                        malla = inscripcion.mi_malla()
                        horas_practica_malla = malla.horas_practicas
                        carrera = inscripcion.carrera.nombre
                        modalidad = inscripcion.modalidad
                        persona = inscripcion.persona
                        cedula = persona.identificacion()
                        nombres = persona.nombre_completo_inverso()
                        correo_inst = persona.emailinst
                        correo_pers = persona.email
                        celular = persona.telefono
                        telefono = persona.telefono_conv
                        canton = persona.canton.nombre
                        practicas = PracticasPreprofesionalesInscripcion.objects.filter(inscripcion=inscripcion, status=True, retirado=False).distinct().order_by('-fecha_creacion')
                        horas_culminadas = null_to_numeric(practicas.filter(~Q(tiposolicitud=3), culminada=True).aggregate(totalhora=Sum('numerohora'))['totalhora'], 0)
                        horas_culminadas_homo = null_to_numeric(practicas.filter(tiposolicitud=3, culminada=True).aggregate(totalhorahomo=Sum('horahomologacion'))['totalhorahomo'], 0)
                        total_horas_practica_culminada = int(horas_culminadas) + int(horas_culminadas_homo)
                        horas_pendientes = 0
                        if horas_practica_malla > 0:
                            horas_pendientes = horas_practica_malla - total_horas_practica_culminada
                        materias_pendientes = inscripcion.cantidad_materias_pendiente_aprobar()
                        ws.write(row_num, 0, u'%s' % malla, formatocelda)
                        ws.write(row_num, 1, u'%s' % facultad, formatocelda)
                        ws.write(row_num, 2, u'%s' % carrera, formatocelda)
                        ws.write(row_num, 3, u'%s' % nivelmalla, formatocelda)
                        ws.write(row_num, 4, u'%s' % modalidad, formatocelda)
                        ws.write(row_num, 5, u'%s' % cedula, formatocelda)
                        ws.write(row_num, 6, u'%s' % nombres, formatocelda)
                        ws.write(row_num, 7, u'%s' % correo_inst, formatocelda)
                        ws.write(row_num, 8, u'%s' % correo_pers, formatocelda)
                        ws.write(row_num, 9, u'%s' % celular, formatocelda)
                        ws.write(row_num, 10, u'%s' % telefono, formatocelda)
                        ws.write(row_num, 11, u'%s' % canton, formatocelda)
                        ws.write(row_num, 12, u'%s' % horas_practica_malla, formatocelda)
                        ws.write(row_num, 13, u'%s' % total_horas_practica_culminada, formatocelda)
                        ws.write(row_num, 14, u'%s' % horas_pendientes, formatocelda)
                        ws.write(row_num, 15, u'%s' % materias_pendientes, formatocelda)
                        ws.write(row_num, 16, u'%s' % "HORAS PRÁCTICAS NO CONFIGURADAS EN LA MALLA" if horas_practica_malla == 0 else "", formatocelda)
                        row_num += 1
                    else:
                        print("No apto ", inscripcion.persona.cedula)
                else:
                    print("No califica ni para revision...", inscripcion.persona.cedula)
            cn.close()
            workbook.close()
            if notificacion_id > 0:
                eNotificacion = Notificacion.objects.get(pk=notificacion_id)
                eNotificacion.en_proceso = False
                eNotificacion.cuerpo = f'Generación del reporte de estudiantes con prácticas pendientes (No Matriculados) {eCoordinacion.nombre.lower()} listo'
                eNotificacion.url = "{}reportes/niveles/{}".format(MEDIA_URL, nombre_archivo)
                eNotificacion.save()
            else:
                eNotificacion = Notificacion(cuerpo=f'Generación del reporte de estudiantes con prácticas pendientes (No Matriculados) {eCoordinacion.nombre.lower()} listo',
                                             titulo=f'Reporte de estudiantes con prácticas pendientes (No Matriculados) {eCoordinacion.alias if eCoordinacion.alias else eCoordinacion.nombre}',
                                             destinatario=ePersona,
                                             url="{}reportes/niveles/{}".format(MEDIA_URL, nombre_archivo),
                                             prioridad=1,
                                             app_label='SGA',
                                             fecha_hora_visible=datetime.now() + timedelta(days=1),
                                             tipo=2,
                                             en_proceso=False)
                eNotificacion.save(request)
            try:
                send_user_notification(user=usernotify,
                                       payload={"head": "Reporte de estudiantes con prácticas pendientes (No Matriculados)",
                                                "body": 'Generación de reporte de estudiantes con prácticas pendientes (No Matriculados) listo',
                                                "action": "notificacion",
                                                "timestamp": time.mktime(datetime.now().timetuple()),
                                                "url": "{}reportes/niveles/{}".format(MEDIA_URL, nombre_archivo),
                                                "btn_notificaciones": traerNotificaciones(request, data, ePersona),
                                                "mensaje": 'Su reporte ha sido generado con exito'
                                                },
                                       ttl=500)
            except:
                pass
        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)
            if notificacion_id > 0:
                eNotificacion = Notificacion.objects.get(pk=notificacion_id)
                eNotificacion.en_proceso = False
                eNotificacion.error = True
                eNotificacion.titulo = f'Reporte de estudiantes con prácticas pendientes (No Matriculados) {eCoordinacion.nombre.lower()} falló en la ejecución'
                eNotificacion.cuerpo = textoerror
                eNotificacion.save()
            else:
                eNotificacion = Notificacion(cuerpo=textoerror,
                                             titulo=f'Reporte de estudiantes con prácticas pendientes (No Matriculados) {eCoordinacion.nombre.lower()} falló en la ejecución',
                                             destinatario=ePersona,
                                             prioridad=1,
                                             app_label='SGA',
                                             fecha_hora_visible=datetime.now() + timedelta(days=1),
                                             tipo=2,
                                             en_proceso=False,
                                             error=True)
                eNotificacion.save(request)
            try:
                send_user_notification(user=usernotify,
                                       payload={"head": "Reporte de estudiantes con prácticas pendientes (No Matriculados) Fallido",
                                                "body": 'Generación del reporte estudiantes con prácticas pendientes (No Matriculados) a fallado',
                                                "action": "notificacion",
                                                "timestamp": time.mktime(datetime.now().timetuple()),
                                                "btn_notificaciones": traerNotificaciones(request, data, ePersona),
                                                "mensaje": textoerror,
                                                "error": True
                                                },
                                       ttl=500)
            except:
                pass


class reporte_estudiantes_porcentaje_asistencia_coordinacion(threading.Thread):

    def __init__(self, request, data, notificacion_id):
        self.request = request
        self.data = data
        self.notificacion_id = notificacion_id
        threading.Thread.__init__(self)

    def run(self):
        from sga.models import Coordinacion, Periodo, Persona, Notificacion, Materia
        from sga.funciones import null_to_numeric
        try:
            directory = os.path.join(MEDIA_ROOT, 'reportes', 'niveles')
            request, data, notificacion_id = self.request, self.data, self.notificacion_id
            usernotify = User.objects.get(pk=request.user.pk)
            ePersona = Persona.objects.get(usuario=usernotify)
            os.makedirs(directory, exist_ok=True)

            try:
                ePeriodo = Periodo.objects.get(pk=int(request.GET.get('idp', '0')))
            except ObjectDoesNotExist:
                raise NameError(u"No se encontro coordinacion académica")
            try:
                eCoordinacion = Coordinacion.objects.get(pk=int(request.GET.get('idc', '0')))
            except ObjectDoesNotExist:
                raise NameError(u"No se encontro coordinacion académica")

            nombre_archivo = "rpt_estudiantes_porcentaje_asistencia_coordinacion_{}_{}.xls".format(eCoordinacion.alias.lower() if eCoordinacion.alias else eCoordinacion.nombre.lower(), random.randint(1, 10000).__str__())
            directory = os.path.join(MEDIA_ROOT, 'reportes', 'niveles', nombre_archivo)
            wb = xlwt.Workbook()
            ws = wb.add_sheet("Hoja1")
            estilo = xlwt.easyxf('font: height 350, name Arial, colour_index black, bold on, italic on; align: wrap on, vert centre, horiz center;')
            ws.write_merge(0, 0, 0, 9, 'UNIVERSIDAD ESTATAL DE MILAGRO', estilo)
            ws.col(1).width = 11000
            ws.col(2).width = 11000
            ws.col(3).width = 27000
            ws.col(4).width = 23000
            ws.col(5).width = 5000
            ws.col(6).width = 3000
            ws.write(4, 1, 'CARRERA')
            ws.write(4, 2, 'PROFESOR')
            ws.write(4, 3, 'MATERIA')
            ws.write(4, 4, 'ALUMNO')
            ws.write(4, 5, 'CEDULA')
            ws.write(4, 6, 'PORCENTAJE')
            a = 4
            for carrera in eCoordinacion.carreras():
                materias = Materia.objects.filter(status=True, nivel__periodo_id=ePeriodo.id, asignaturamalla__malla__carrera=carrera).exclude(nivel__modalidad_id=3)
                for mat in materias:
                    for asignadomateria in mat.asignados_a_esta_materia():
                        por = 0
                        if asignadomateria.asistencias_zoom_valida() > 0:
                            por = round(((asignadomateria.asistencias_zoom_valida() * 100) / asignadomateria.cantidad_asistencias_zoom()), 0)
                            if por < 70:
                                a += 1
                                ws.write(a, 1, mat.asignaturamalla.malla.carrera.__str__())
                                ws.write(a, 2, str(mat.profesores_materia()[0].profesor))
                                ws.write(a, 3, mat.nombre_completo())
                                ws.write(a, 4, asignadomateria.matricula.inscripcion.persona.nombre_completo())
                                ws.write(a, 5, asignadomateria.matricula.inscripcion.persona.identificacion())
                                ws.write(a, 6, por)
            wb.save(directory)
            if notificacion_id > 0:
                eNotificacion = Notificacion.objects.get(pk=notificacion_id)
                eNotificacion.en_proceso = False
                eNotificacion.cuerpo = f'Generación del reporte de estudiantes porcentaje asistencia {eCoordinacion.nombre.lower()} listo'
                eNotificacion.url = "{}reportes/niveles/{}".format(MEDIA_URL, nombre_archivo)
                eNotificacion.save()
            else:
                eNotificacion = Notificacion(cuerpo=f'Generación del reporte de estudiantes porcentaje asistencia {eCoordinacion.nombre.lower()} listo',
                                             titulo=f'Reporte de estudiantes porcentaje asistencia{eCoordinacion.alias if eCoordinacion.alias else eCoordinacion.nombre}',
                                             destinatario=ePersona,
                                             url="{}reportes/niveles/{}".format(MEDIA_URL, nombre_archivo),
                                             prioridad=1,
                                             app_label='SGA',
                                             fecha_hora_visible=datetime.now() + timedelta(days=1),
                                             tipo=2,
                                             en_proceso=False)
                eNotificacion.save(request)
            try:
                send_user_notification(user=usernotify,
                                       payload={"head": "Reporte de estudiantes porcentaje asistencia",
                                                "body": 'Generación de reporte de estudiantes porcentaje asistencia listo',
                                                "action": "notificacion",
                                                "timestamp": time.mktime(datetime.now().timetuple()),
                                                "url": "{}reportes/niveles/{}".format(MEDIA_URL, nombre_archivo),
                                                "btn_notificaciones": traerNotificaciones(request, data, ePersona),
                                                "mensaje": 'Su reporte ha sido generado con exito'
                                                },
                                       ttl=500)
            except:
                pass
        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)
            if notificacion_id > 0:
                eNotificacion = Notificacion.objects.get(pk=notificacion_id)
                eNotificacion.en_proceso = False
                eNotificacion.error = True
                eNotificacion.titulo = f'Reporte de estudiantes porcentaje asistencia {eCoordinacion.nombre.lower()} falló en la ejecución'
                eNotificacion.cuerpo = textoerror
                eNotificacion.save()
            else:
                eNotificacion = Notificacion(cuerpo=textoerror,
                                             titulo=f'Reporte de estudiantes porcentaje asistencia {eCoordinacion.nombre.lower()} falló en la ejecución',
                                             destinatario=ePersona,
                                             prioridad=1,
                                             app_label='SGA',
                                             fecha_hora_visible=datetime.now() + timedelta(days=1),
                                             tipo=2,
                                             en_proceso=False,
                                             error=True)
                eNotificacion.save(request)
            try:
                send_user_notification(user=usernotify,
                                       payload={"head": "Reporte de estudiantes porcentaje asistencia Fallido",
                                                "body": 'Generación del reporte estudiantes porcentaje asistencia a fallado',
                                                "action": "notificacion",
                                                "timestamp": time.mktime(datetime.now().timetuple()),
                                                "btn_notificaciones": traerNotificaciones(request, data, ePersona),
                                                "mensaje": textoerror,
                                                "error": True
                                                },
                                       ttl=500)
            except:
                pass


class informe_conflicto_horario_alumno(threading.Thread):

    def __init__(self, request, data, notificacion_id):
        self.request = request
        self.data = data
        self.notificacion_id = notificacion_id
        threading.Thread.__init__(self)

    def run(self):
        from sga.commonviews import conflicto_materias_estudiante
        from sga.funcionesxhtml2pdf import convert_html_to_pdf
        from sga.models import Nivel, Materia, Persona, Notificacion, AlumnosPracticaMateria, Matricula
        try:
            request, data, notificacion_id = self.request, self.data, self.notificacion_id
            usernotify = User.objects.get(pk=request.user.pk)
            ePersona = Persona.objects.get(usuario=usernotify)
            directory = os.path.join(MEDIA_ROOT, 'reportes', 'niveles')
            os.makedirs(directory, exist_ok=True)
            nombre_archivo = "informe_conflicto_estudiante{}.pdf".format(random.randint(1, 10000).__str__())
            try:
                eNivel = Nivel.objects.get(pk=int(request.GET.get('idn', '0')))
            except ObjectDoesNotExist:
                raise NameError(u"No se encontro nivel académico")
                # VERIFICAR CONFLICTO ALUMNOS
            listaconflicto = []
            ePeriodo = eNivel.periodo
            eMatriculas = Matricula.objects.filter(status=True, nivel=eNivel, nivel__periodo=ePeriodo, retiradomatricula=False).order_by('inscripcion__persona')
            for eMatricula in eMatriculas:
                eMaterias = Materia.objects.filter(id__in=eMatricula.materiaasignada_set.values_list('materia__id').filter(status=True, sinasistencia=False), status=True)
                alumnaspracticascongrupo = AlumnosPracticaMateria.objects.values_list('profesormateria__id', 'grupoprofesor__id').filter(materiaasignada__materia__id__in=eMaterias.values_list('id'), materiaasignada__matricula=eMatricula, grupoprofesor__isnull=False)
                conflicto = conflicto_materias_estudiante(materias=eMaterias, lista_pm_grupo=alumnaspracticascongrupo, extraerlistaclasesconflicto=True)
                if conflicto:
                    listaconflicto.append([eMatricula, conflicto[0], conflicto[1]])
            valida = convert_html_to_pdf('niveles/informeconflictoestudiante.html',
                                         {'pagesize': 'A4 landscape', 'periodo': ePeriodo,
                                          'listaconflicto': listaconflicto, 'nivel': eNivel},
                                         nombre_archivo,
                                         directory
                                         )
            if not valida:
                raise NameError(u"Informe con error en la generación")

            if notificacion_id > 0:
                eNotificacion = Notificacion.objects.get(pk=notificacion_id)
                eNotificacion.en_proceso = False
                eNotificacion.cuerpo = 'Generación de informe conflicto horario estudiante listo'
                eNotificacion.url = "{}reportes/niveles/{}".format(MEDIA_URL, nombre_archivo)
                eNotificacion.save()
            else:
                eNotificacion = Notificacion(cuerpo='Generación de informe conflicto horario estudiante listo',
                                             titulo='Informe de conflicto de horario estudiante',
                                             destinatario=ePersona,
                                             url="{}reportes/niveles/{}".format(MEDIA_URL, nombre_archivo),
                                             prioridad=1,
                                             app_label='SGA',
                                             fecha_hora_visible=datetime.now() + timedelta(days=1),
                                             tipo=2,
                                             en_proceso=False)
                eNotificacion.save(request)
            try:
                send_user_notification(user=usernotify,
                                       payload={"head": "Informe de conflicto de horario estudiante",
                                                "body": 'Generación de informe conflicto horario estudiante listo',
                                                "action": "notificacion",
                                                "timestamp": time.mktime(datetime.now().timetuple()),
                                                "url": "{}reportes/niveles/{}".format(MEDIA_URL, nombre_archivo),
                                                "btn_notificaciones": traerNotificaciones(request, data, ePersona),
                                                "mensaje": 'Su reporte ha sido generado con exito'
                                                },
                                       ttl=500)
            except:
                pass
        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)
            if notificacion_id > 0:
                eNotificacion = Notificacion.objects.get(pk=notificacion_id)
                eNotificacion.en_proceso = False
                eNotificacion.error = True
                eNotificacion.titulo = 'Informe de conflicto de horario estudiante falló en la ejecución'
                eNotificacion.cuerpo = textoerror
                eNotificacion.save()
            else:
                eNotificacion = Notificacion(cuerpo=textoerror,
                                             titulo='Informe de conflicto de horario estudiante falló en la ejecución',
                                             destinatario=ePersona,
                                             prioridad=1,
                                             app_label='SGA',
                                             fecha_hora_visible=datetime.now() + timedelta(days=1),
                                             tipo=2,
                                             en_proceso=False,
                                             error=True)
                eNotificacion.save(request)
            try:
                send_user_notification(user=usernotify,
                                       payload={"head": "Informe de conflicto de horario estudiante Fallido",
                                                "body": 'Generación de informe conflicto horario estudiante a fallado',
                                                "action": "notificacion",
                                                "timestamp": time.mktime(datetime.now().timetuple()),
                                                "btn_notificaciones": traerNotificaciones(request, data, ePersona),
                                                "mensaje": textoerror,
                                                "error": True
                                                },
                                       ttl=500)
            except:
                pass


class reporte_califcacion_periodo(threading.Thread):

    def __init__(self, request, data, periodo_id, modalidad_id, notificacion_id):
        self.request = request
        self.data = data
        self.periodo_id = periodo_id
        self.modalidad_id = modalidad_id
        self.notificacion_id = notificacion_id
        threading.Thread.__init__(self)

    def run(self):
        from sga.models import Coordinacion, Periodo, Persona, Notificacion, Materia, Modalidad, Malla, \
            MateriaAsignada, ModeloEvaluativo, DetalleModeloEvaluativo
        from sga.funciones import null_to_numeric, null_to_decimal

        try:
            directory = os.path.join(MEDIA_ROOT, 'reportes', 'calificaciones', 'periodos', 'alumnos')
            request, data, periodo_id, modalidad_id, notificacion_id = self.request, self.data, self.periodo_id, self.modalidad_id, self.notificacion_id
            os.makedirs(directory, exist_ok=True)

            try:
                ePeriodo = Periodo.objects.get(pk=periodo_id)
            except ObjectDoesNotExist:
                raise NameError(u"No se encontro periodo académico")
            try:
                eModalidad = Modalidad.objects.get(pk=modalidad_id)
            except ObjectDoesNotExist:
                raise NameError(u"No se encontro modalidad")
            nombre_archivo = "rpt_calificaciones_idp_{}_{}.xlsx".format(ePeriodo.pk, random.randint(1, 10000).__str__())
            directory = os.path.join(MEDIA_ROOT, 'reportes', 'calificaciones', 'periodos', 'alumnos', nombre_archivo)
            usernotify = User.objects.get(pk=request.user.pk)
            ePersona = Persona.objects.get(usuario=usernotify)
            eMallasIngles = Malla.objects.filter(pk__in=[353, 22]).values_list('id', flat=True)
            eMaterias = Materia.objects.filter(nivel__periodo=ePeriodo, status=True, asignaturamalla__malla__modalidad=eModalidad).exclude(Q(asignatura_id=4837) | Q(materiaasignada__materia__asignaturamalla__malla_id__in=eMallasIngles.values_list('id', flat=True)))
            eMateriaAsignadas = MateriaAsignada.objects.filter(materia__in=eMaterias, matricula__status=True, matricula__bloqueomatricula=False, retiramateria=False, status=True)
            eMateriaAsignadas = eMateriaAsignadas.order_by('matricula__inscripcion__persona__apellido1', 'matricula__inscripcion__persona__apellido2', 'matricula__inscripcion__persona__nombres').distinct()
            if DEBUG:
                eMateriaAsignadas = eMateriaAsignadas[0:10]
            eModeloEvaluativos = ModeloEvaluativo.objects.filter(pk__in=eMaterias.values_list("modeloevaluativo__id", flat=True))
            eDetalleModeloEvaluativos = DetalleModeloEvaluativo.objects.filter(modelo__in=eModeloEvaluativos).order_by('modelo', 'orden')
            __author__ = 'Unemi'
            workbook = xlsxwriter.Workbook(directory, {'constant_memory': True})
            ws = workbook.add_worksheet('calificaciones')
            title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
            titulo2 = easyxf('font: name Times New Roman, color-index black, bold on , height 250; alignment: horiz centre')
            font_style = XFStyle()
            font_style.font.bold = True
            font_style2 = XFStyle()
            font_style2.font.bold = False
            fuentetitulo = workbook.add_format({'align': 'center',
                                                'valign': 'vcenter',
                                                'bold': 1,
                                                'text_wrap': True,
                                                'font_size': 16})
            fuentesubtitulo = workbook.add_format({'align': 'center',
                                                   'valign': 'vcenter',
                                                   'bold': 1,
                                                   'text_wrap': True,
                                                   'font_size': 14})
            fuentecabecera = workbook.add_format({'align': 'center',
                                                  'valign': 'vcenter',
                                                  'bold': 1,
                                                  'border': 1,
                                                  'text_wrap': True,
                                                  'fg_color': '#1C3247',
                                                  'font_color': 'white'})
            formatoceldacenter = workbook.add_format({'border': 1,
                                                      'valign': 'vcenter',
                                                      'align': 'center',
                                                      'text_wrap': True
                                                      })
            formatoceldaleft = workbook.add_format({'border': 1,
                                                    'valign': 'vcenter',
                                                    'align': 'left',
                                                    'text_wrap': True
                                                    })
            formatoceldadecimal = workbook.add_format({'num_format': '$ #,##0,00',
                                                       'text_wrap': True,
                                                       'align': 'center',
                                                       'valign': 'vcenter',
                                                       'border': 1})


            columns = [
                (u"#", 10),
                (u"TIPO DOCUMENTO", 20),
                (u"DOCUMENTO", 20),
                (u"ALUMNO", 50),
                (u"FACULTAD", 80),
                (u"CARRERA", 80),
                (u"MODALIDAD", 20),
                (u"NIVEL", 20),
                (u"ASIGNATURA", 150),
                (u"PARALELO", 20),
                (u"PROFESOR", 50),
            ]
            aDetalleModeloEvaluativos = []
            for eDetalleModeloEvaluativo in eDetalleModeloEvaluativos:
                columns.append((f"{eDetalleModeloEvaluativo.nombre}", 70))
                aDetalleModeloEvaluativos.append({'id': eDetalleModeloEvaluativo.id,  'name': eDetalleModeloEvaluativo.nombre})
            row_num = 3
            for col_num in range(len(columns)):
                ws.write(row_num, col_num, columns[col_num][0], fuentecabecera)
                ws.set_column(row_num, col_num, columns[col_num][1])
            col_num = len(columns)
            ws.write(row_num, col_num, "NOTA FINAL", fuentecabecera)
            ws.set_column(row_num, col_num, 10)
            ws.merge_range(1, 1, 1, col_num + 1, "UNIVERSIDAD ESTATAL DE MILAGRO", fuentetitulo)
            ws.merge_range(2, 1, 2, col_num + 1, f'Calificaciones del período académico {ePeriodo.nombre} de carreras en modalidad {eModalidad.nombre}', fuentesubtitulo)
            row_num = 4
            i = 0
            for eMateriaAsignada in eMateriaAsignadas:
                i += 1
                eMateria = eMateriaAsignada.materia
                eModeloEvaluativo = eMateria.modeloevaluativo
                eDetalleModeloEvaluativos = eModeloEvaluativo.campos()
                eParalelo = eMateria.paralelomateria
                eProfesor = None
                if eMateria.asignaturamalla.malla.modalidad.es_enlinea():
                    eProfesor = eMateria.profesor_principal_virtual()
                elif eMateria.asignaturamalla.malla.modalidad.es_presencial() or eMateria.asignaturamalla.malla.modalidad.es_semipresencial():
                    eProfesor = eMateria.profesor_principal()
                eAsignaturaMalla = eMateria.asignaturamalla
                eMalla = eAsignaturaMalla.malla
                eNivelMalla = eAsignaturaMalla.nivelmalla
                eAsignatura = eAsignaturaMalla.asignatura
                eMatricula = eMateriaAsignada.matricula
                eInscripcion = eMatricula.inscripcion
                eModalidad = eMalla.modalidad
                eFacultad = eInscripcion.coordinacion
                eCarrera = eInscripcion.carrera
                ePersona = eInscripcion.persona
                eEvaluacionGenericas = eMateriaAsignada.evaluacion_generica()

                ws.write(row_num, 0, u"%s" % i, formatoceldacenter)
                ws.set_column(row_num, 0, 10)

                ws.write(row_num, 1, u"%s" % ePersona.tipo_documento(), formatoceldacenter)
                ws.set_column(row_num, 1, 20)

                ws.write(row_num, 2, u"%s" % ePersona.documento(), formatoceldacenter)
                ws.set_column(row_num, 2, 20)

                ws.write(row_num, 3, u"%s" % ePersona.nombre_completo(), formatoceldaleft)
                ws.set_column(row_num, 3, 50)

                ws.write(row_num, 4, u"%s" % eFacultad.nombre if eFacultad else '', formatoceldaleft)
                ws.set_column(row_num, 4, 80)

                ws.write(row_num, 5, u"%s" % eCarrera.nombrevisualizar if eCarrera.nombrevisualizar else eCarrera.nombre, formatoceldaleft)
                ws.set_column(row_num, 5, 80)

                ws.write(row_num, 6, u"%s" % eModalidad.nombre if eModalidad else '', formatoceldacenter)
                ws.set_column(row_num, 6, 20)

                ws.write(row_num, 7, u"%s" % eNivelMalla.nombre, formatoceldacenter)
                ws.set_column(row_num, 7, 20)

                ws.write(row_num, 8, u"%s" % eAsignatura.nombre, formatoceldaleft)
                ws.set_column(row_num, 8, 150)

                ws.write(row_num, 9, u"%s" % eParalelo.nombre if eParalelo else '', formatoceldacenter)
                ws.set_column(row_num, 9, 20)

                ws.write(row_num, 10, u"%s" % eProfesor.persona.nombre_completo() if eProfesor else '', formatoceldaleft)
                ws.set_column(row_num, 10, 50)

                col_num = 10
                for aDetalleModeloEvaluativo in aDetalleModeloEvaluativos:
                    id = aDetalleModeloEvaluativo.get('id', 0)
                    col_num += 1
                    if (eDetalleModeloEvaluativo := eDetalleModeloEvaluativos.filter(status=True, pk=id).first()) is not None:
                        if (eCalificacion := eEvaluacionGenericas.filter(detallemodeloevaluativo=eDetalleModeloEvaluativo).first()) is not None:
                            ws.write(row_num, col_num, u"%s" % null_to_decimal(eCalificacion.valor, eCalificacion.detallemodeloevaluativo.decimales), formatoceldadecimal)
                            ws.set_column(row_num, col_num, 10)
                        else:
                            ws.write(row_num, col_num, "0.0", formatoceldadecimal)
                            ws.set_column(row_num, col_num, 10)
                    else:
                        ws.write(row_num, col_num, "-", formatoceldadecimal)
                        ws.set_column(row_num, col_num, 10)
                col_num += 1
                ws.write(row_num, col_num, u"%s" % null_to_decimal(eMateriaAsignada.notafinal, 2), formatoceldadecimal)
                ws.set_column(row_num, col_num, 10)
                row_num += 1
            workbook.close()

            if notificacion_id > 0:
                eNotificacion = Notificacion.objects.get(pk=notificacion_id)
                eNotificacion.en_proceso = False
                eNotificacion.cuerpo = f'Generación del reporte de califcaciones del período académico {ePeriodo.nombre} de carreras en modalidad {eModalidad.nombre} listo'
                eNotificacion.url = "{}reportes/calificaciones/periodos/alumnos/{}".format(MEDIA_URL, nombre_archivo)
                eNotificacion.save(request)
            else:
                eNotificacion = Notificacion(cuerpo=f'Generación del reporte de califcaciones del período académico {ePeriodo.nombre} de carreras en modalidad {eModalidad.nombre} listo',
                                             titulo=f'Reporte de calificaciones del período académico {ePeriodo.nombre} de carreras en modalidad {eModalidad.nombre}',
                                             destinatario=ePersona,
                                             url="{}reportes/calificaciones/periodos/alumnos/{}".format(MEDIA_URL, nombre_archivo),
                                             prioridad=1,
                                             app_label='SGA',
                                             fecha_hora_visible=datetime.now() + timedelta(days=1),
                                             tipo=2,
                                             en_proceso=False)
                eNotificacion.save(request)
            try:
                send_user_notification(user=usernotify,
                                       payload={"head": f'Reporte de calificaciones del período académico {ePeriodo.nombre} de carreras en modalidad {eModalidad.nombre}',
                                                "body": f'Generación del reporte de califcaciones del período académico {ePeriodo.nombre} de carreras en modalidad {eModalidad.nombre} listo',
                                                "action": "notificacion",
                                                "timestamp": time.mktime(datetime.now().timetuple()),
                                                "url": "{}reportes/calificaciones/periodos/alumnos/{}".format(MEDIA_URL, nombre_archivo),
                                                "btn_notificaciones": traerNotificaciones(request, data, ePersona),
                                                "mensaje": 'Su reporte ha sido generado con exito'
                                                },
                                       ttl=500)
            except:
                pass
        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)
            if notificacion_id > 0:
                eNotificacion = Notificacion.objects.get(pk=notificacion_id)
                eNotificacion.titulo=f'Reporte de calificaciones del período académico {ePeriodo.nombre} de carreras en modalidad {eModalidad.nombre} falló en la ejecución'
                eNotificacion.en_proceso = False
                eNotificacion.cuerpo = textoerror
                eNotificacion.url = None
                eNotificacion.save(request)
            else:
                eNotificacion = Notificacion(cuerpo=textoerror,
                                             titulo=f'Reporte de calificaciones del período académico {ePeriodo.nombre} de carreras en modalidad {eModalidad.nombre} falló en la ejecución',
                                             destinatario=ePersona,
                                             url=None,
                                             prioridad=1,
                                             app_label='SGA',
                                             fecha_hora_visible=datetime.now() + timedelta(days=1),
                                             tipo=2,
                                             en_proceso=False)
                eNotificacion.save(request)
            try:
                send_user_notification(user=usernotify,
                                       payload={"head": f'Reporte de calificaciones del período académico {ePeriodo.nombre} de carreras en modalidad {eModalidad.nombre} fallido',
                                                "body": f'Generación del reporte de califcaciones del período académico {ePeriodo.nombre} de carreras en modalidad {eModalidad.nombre} a fallado',
                                                "action": "notificacion",
                                                "timestamp": time.mktime(datetime.now().timetuple()),
                                                "btn_notificaciones": traerNotificaciones(request, data, ePersona),
                                                "mensaje": textoerror,
                                                "error": True
                                                },
                                       ttl=500)
            except:
                pass


class reporte_ultimo_modulo_ingles_aprobado(threading.Thread):

    def __init__(self, request, data, periodo_id=0, aplica=False, coordinacion_id=None, notificacion_id=0):
        self.request = request
        self.data = data
        self.periodo_id = periodo_id
        self.aplica = aplica
        self.coordinacion_id = coordinacion_id
        self.notificacion_id = notificacion_id
        threading.Thread.__init__(self)

    def run(self):
        from sga.models import Coordinacion, Periodo, Persona, Notificacion, Matricula, RecordAcademico
        from sga.funciones import null_to_numeric, null_to_decimal

        try:
            directory = os.path.join(MEDIA_ROOT, 'reportes', 'ingles', 'alumnos')
            request, data, periodo_id, aplica, coordinacion_id, notificacion_id = self.request, self.data, self.periodo_id, self.aplica, self.coordinacion_id, self.notificacion_id
            os.makedirs(directory, exist_ok=True)
            try:
                ePeriodo = Periodo.objects.get(pk=periodo_id)
            except ObjectDoesNotExist:
                raise NameError(u"No se encontro un periodo válido")

            if coordinacion_id > 0:
                eCoordinaciones = Coordinacion.objects.filter(pk=coordinacion_id)
                nombre_archivo = "rpt_estudiantes_ultimo_modulo_ingles_aprobado_{}_{}.xlsx".format(eCoordinaciones.first().alias, random.randint(1, 10000).__str__())
                coordinacion_nombre = f"{eCoordinaciones.first().nombre}"
            else:
                eCoordinaciones = Coordinacion.objects.filter(pk__in=[1, 2, 3, 4, 5])
                nombre_archivo = "rpt_estudiantes_ultimo_modulo_ingles_aprobado_todas_{}.xlsx".format(random.randint(1, 10000).__str__())
                coordinacion_nombre = f"Todas las facultades"
            directory = os.path.join(MEDIA_ROOT, 'reportes', 'ingles', 'alumnos', nombre_archivo)
            usernotify = User.objects.get(pk=request.user.pk)
            ePersona = Persona.objects.get(usuario=usernotify)
            __author__ = 'Unemi'

            workbook = xlsxwriter.Workbook(directory, {'constant_memory': True})
            ws = workbook.add_worksheet('Hoja1')
            ws.set_column(0, 10, 30)
            ws.set_column(1, 10, 30)
            formatotitulo = workbook.add_format({'bold': 1, 'text_wrap': True, 'border': 1, 'align': 'center', 'valign': 'middle', 'fg_color': '#A2D0EC'})
            formatotitulo_filtros = workbook.add_format({'bold': 1, 'text_wrap': True, 'border': 1, 'fg_color': '#EBF5FB'})
            formatoceldacab = workbook.add_format({'align': 'center', 'bold': 1, 'border': 1, 'text_wrap': True, 'fg_color': '#EBF5FB'})
            formatoceldaleft = workbook.add_format({'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1})
            ws.merge_range('A1:J1', F'ESTUDIANTES DEL PERIODO {ePeriodo.nombre} CON ÚLTIMO MÓDULO DE INGLÉS APROBADO', formatotitulo)
            ws.merge_range('A2:J2', str(coordinacion_nombre), formatotitulo)
            ws.write('A3', 'INSCRIPCIÓN', formatoceldacab)
            ws.write('B3', 'MATRÍCULA', formatoceldacab)
            ws.write('C3', 'IDENTIFICACIÓN', formatoceldacab)
            ws.write('D3', 'NOMBRES', formatoceldacab)
            ws.write('E3', 'EMAIL PERSONAL', formatoceldacab)
            ws.write('F3', 'EMAIL INSTITUCIONAL', formatoceldacab)
            ws.write('G3', 'TELÉFONO', formatoceldacab)
            ws.write('H3', 'CARRERA', formatoceldacab)
            ws.write('I3', 'MODALIDAD', formatoceldacab)
            ws.write('J3', 'ÚLTIMO MÓDULO APROBADO', formatoceldacab)
            eMatriculas = Matricula.objects.filter(status=True, retiradomatricula=False, nivel__periodo=ePeriodo, inscripcion__coordinacion__in=eCoordinaciones).order_by('-inscripcion__carrera')
            if DEBUG:
                eMatriculas = eMatriculas[:100]
            fila_det = 4
            for eMatricula in eMatriculas:
                eInscripcion = eMatricula.inscripcion
                eMalla = eInscripcion.mi_malla()
                ws.write('A%s' % fila_det, str(eInscripcion.id if eInscripcion else 'NO EXISTE REGISTRO'), formatoceldaleft)
                ws.write('B%s' % fila_det, str(eMatricula.id if eMatricula.id else 'NO EXISTE REGISTRO'), formatoceldaleft)
                ws.write('C%s' % fila_det, str(eInscripcion.persona.documento() if eInscripcion else 'NO EXISTE REGISTRO'), formatoceldaleft)
                ws.write('D%s' % fila_det, str(eInscripcion.persona.nombre_completo_inverso() if eInscripcion.persona else 'NO EXISTE REGISTRO'), formatoceldaleft)
                ws.write('E%s' % fila_det, str(eInscripcion.persona.email if eInscripcion.persona else 'NO EXISTE REGISTRO'), formatoceldaleft)
                ws.write('F%s' % fila_det, str(eInscripcion.persona.emailinst if eInscripcion.persona else 'NO EXISTE REGISTRO'), formatoceldaleft)
                ws.write('G%s' % fila_det, str(eInscripcion.persona.telefono if eInscripcion.persona else 'NO EXISTE REGISTRO'), formatoceldaleft)
                if eMalla:
                    ws.write('H%s' % fila_det, str(eMalla.carrera.nombre), formatoceldaleft)
                    ws.write('I%s' % fila_det, str(eMalla.modalidad.nombre if eMalla.modalidad else 'NO EXISTE REGISTRO'), formatoceldaleft)
                else:
                    ws.write('H%s' % fila_det, str(eInscripcion.carrera.nombre if eInscripcion.carrera else 'NO EXISTE REGISTRO'), formatoceldaleft)
                    ws.write('I%s' % fila_det, str(eInscripcion.carrera.get_modalidad_display() if eInscripcion.carrera else 'NO EXISTE REGISTRO'), formatoceldaleft)
                eModulo = None
                if (last_modulo := RecordAcademico.objects.filter(status=True, inscripcion=eMatricula.inscripcion, aprobada=True, modulomalla__isnull=False).order_by('-modulomalla__orden').distinct().first()) is not None:
                    eModulo = last_modulo.modulomalla
                ws.write('J%s' % fila_det, str(eModulo if eModulo else 'NO REGISTRA'), formatoceldaleft)
                fila_det += 1

            workbook.close()

            if notificacion_id > 0:
                eNotificacion = Notificacion.objects.get(pk=notificacion_id)
                eNotificacion.en_proceso = False
                eNotificacion.cuerpo = f'Generación del reporte de estudiantes con último módulo de inglés aprobado del período académico {ePeriodo.nombre} listo'
                eNotificacion.url = "{}reportes/ingles/alumnos/{}".format(MEDIA_URL, nombre_archivo)
                eNotificacion.save(request)
            else:
                eNotificacion = Notificacion(cuerpo=f'Generación del reporte de estudiantes con último módulo de inglés aprobado del período académico {ePeriodo.nombre} listo',
                                             titulo=f'Reporte de estudiantes con último módulo de inglés aprobado del período académico {ePeriodo.nombre}',
                                             destinatario=ePersona,
                                             url="{}reportes/ingles/alumnos/{}".format(MEDIA_URL, nombre_archivo),
                                             prioridad=1,
                                             app_label='SGA',
                                             fecha_hora_visible=datetime.now() + timedelta(days=1),
                                             tipo=2,
                                             en_proceso=False)
                eNotificacion.save(request)
            try:
                send_user_notification(user=usernotify,
                                       payload={"head": f'Reporte de estudiantes con último módulo de inglés aprobado',
                                                "body": f'Generación del reporte de estudiantes con último módulo de inglés aprobado listo',
                                                "action": "notificacion",
                                                "timestamp": time.mktime(datetime.now().timetuple()),
                                                "url": "{}reportes/ingles/alumnos/{}".format(MEDIA_URL, nombre_archivo),
                                                "btn_notificaciones": traerNotificaciones(request, data, ePersona),
                                                "mensaje": 'Su reporte ha sido generado con exito'
                                                },
                                       ttl=500)
            except:
                pass
        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)
            if notificacion_id > 0:
                eNotificacion = Notificacion.objects.get(pk=notificacion_id)
                eNotificacion.titulo=f'Reporte de estudiantes con último módulo de inglés aprobado del período académico {ePeriodo.nombre} falló en la ejecución'
                eNotificacion.en_proceso = False
                eNotificacion.cuerpo = textoerror
                eNotificacion.url = None
                eNotificacion.save(request)
            else:
                eNotificacion = Notificacion(cuerpo=textoerror,
                                             titulo=f'Reporte de estudiantes con último módulo de inglés aprobado del período académico {ePeriodo.nombre} falló en la ejecución',
                                             destinatario=ePersona,
                                             url=None,
                                             prioridad=1,
                                             app_label='SGA',
                                             fecha_hora_visible=datetime.now() + timedelta(days=1),
                                             tipo=2,
                                             en_proceso=False)
                eNotificacion.save(request)
            try:
                send_user_notification(user=usernotify,
                                       payload={"head": f'Reporte de estudiantes con último módulo de inglés aprobado fallido',
                                                "body": f'Generación del reporte de estudiantes con último módulo de inglés aprobado a fallado',
                                                "action": "notificacion",
                                                "timestamp": time.mktime(datetime.now().timetuple()),
                                                "btn_notificaciones": traerNotificaciones(request, data, ePersona),
                                                "mensaje": textoerror,
                                                "error": True
                                                },
                                       ttl=500)
            except:
                pass


class reporte_modulos_ingles_aprobado(threading.Thread):

    def __init__(self, request, data, periodo_id, notificacion_id):
        self.request = request
        self.data = data
        self.periodo_id = periodo_id
        self.notificacion_id = notificacion_id
        threading.Thread.__init__(self)

    def run(self):
        from sga.models import Coordinacion, Periodo, Persona, Notificacion, Matricula, RecordAcademico
        from sga.funciones import null_to_numeric, null_to_decimal

        try:
            directory = os.path.join(MEDIA_ROOT, 'reportes', 'ingles', 'alumnos')
            request, data, periodo_id, notificacion_id = self.request, self.data, self.periodo_id, self.notificacion_id
            os.makedirs(directory, exist_ok=True)
            try:
                ePeriodo = Periodo.objects.get(pk=periodo_id)
            except ObjectDoesNotExist:
                raise NameError(u"No se encontro un periodo válido")

            nombre_archivo = "rpt_estudiantes_modulo_ingles_aprobado_{}.xlsx".format(random.randint(1, 10000).__str__())
            directory = os.path.join(MEDIA_ROOT, 'reportes', 'ingles', 'alumnos', nombre_archivo)
            usernotify = User.objects.get(pk=request.user.pk)
            ePersona = Persona.objects.get(usuario=usernotify)
            __author__ = 'Unemi'
            workbook = xlsxwriter.Workbook(directory, {'constant_memory': True})
            ws = workbook.add_worksheet('Hoja1')
            formatocabeceracolumna = workbook.add_format({'bold': 1, 'border': 1, 'align': 'center', 'valign': 'vcenter', 'bg_color': 'silver', 'text_wrap': 1, 'font_size': 10})
            formatocelda = workbook.add_format({'border': 1})
            formatoceldaentero = workbook.add_format({'border': 1, 'num_format': '0'})
            formatotitulo = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'bold': 1, 'font_size': 14})
            ws.merge_range(0, 0, 0, 6, 'UNIVERSIDAD ESTATAL DE MILAGRO', formatotitulo)
            ws.merge_range(1, 0, 1, 6, 'LISTADO DE ESTUDIANTES CON MÓDULOS DE INGLÉS APROBADOS', formatotitulo)
            ws.merge_range(2, 0, 2, 6, 'PERIODO ' + ePeriodo.nombre, formatotitulo)
            columns = [
                (u"CÉDULA", 15),
                (u"ESTUDIANTE", 40),
                (u"CORREO PERSONAL", 40),
                (u"CORREO INSTITUCIONAL", 40),
                (u"TELÉFONO", 40),
                (u"CARRERA", 40),
                (u"MODALIDAD", 20),
                (u"NIVEL MATRICULA", 20),
                (u"AÑO MALLA", 10),
                (u"MÓDULOS APROBADOS", 20),
                (u"MÓDULOS REPROBADOS", 20),
                (u"CÓDIGO CARRERA", 20),
                (u"CÓDIGO MÓDULO", 20),
                (u"ID MATRICULA", 20)
            ]

            row_num = 4
            for col_num in range(len(columns)):
                ws.write(row_num, col_num, columns[col_num][0], formatocabeceracolumna)
                ws.set_column(col_num, col_num, columns[col_num][1])

            cursor = connections['sga_select'].cursor()
            sql = """
                        Select pe.nombre as periodo, per.cedula, (per.apellido1 || ' ' || per.apellido2 || ' ' || per.nombres) as estudiante,
                         per.email,per.emailinst,car.nombre as carrera, mo.nombre as modalidad, nm.nombre as nivelmatricula,
                        extract (year from malla.inicio) as aniomalla,
                        (Select count(r.id) From
                        sga_recordacademico r  
                        Inner join sga_malla malla2 on malla2.id=22
                        Inner join sga_asignaturamalla asi on asi.malla_id=malla2.id  
                        Where  r.inscripcion_id=i.id  
                        and r.aprobada=True and r.asignatura_id=asi.asignatura_id
                        and r.asignatura_id<>782) as modulosinglesaprobados,
                        (Select count(reco.id) From
                        sga_recordacademico reco  
                        Inner join sga_malla malla2 on malla2.id=22
                        Inner join sga_asignaturamalla asi on asi.malla_id=malla2.id  
                        Where  reco.inscripcion_id=i.id  
                        and reco.aprobada= FALSE and reco.asignatura_id=asi.asignatura_id
                        and reco.asignatura_id<>782) as modulosinglesreprobados, car.codigo AS cod,
                        per.telefono,mat.id
                        From sga_matricula mat
                        Inner join sga_nivel ni on mat.nivel_id=ni.id
                        Inner join sga_inscripcion i on i.id=mat.inscripcion_id
                        Inner join sga_persona per on per.id=i.persona_id
                        Inner join sga_carrera car on car.id=i.carrera_id
                        Inner join sga_modalidad mo on mo.id=car.modalidad
                        Inner join sga_nivelmalla nm on nm.id=mat.nivelmalla_id
                        Inner join sga_inscripcionmalla im on im.inscripcion_id=i.id
                        Inner join sga_malla malla on malla.id=im.malla_id
                        Inner join sga_periodo pe on pe.id=ni.periodo_id
                        Where pe.id=%s AND i.activo=TRUE AND mat.retiradomatricula=FALSE AND i.id NOT IN (
                                SELECT ret.inscripcion_id
                                FROM sga_retirocarrera ret WHERE ret.status
                                )
                        Order By estudiante asc 
                    """ % ePeriodo.pk

            cursor.execute(sql)
            results = cursor.fetchall()

            row_num = 5
            for r in results:
                cedula = r[1]
                estudiante = r[2]
                emailper = r[3]
                emailinst = r[4]
                carrera = r[5]
                telefono = r[12]
                id_matricula = r[13]
                modalidad = r[6]
                nivelmatricula = r[7]
                aniomalla = r[8]
                totalaprobado = r[9]
                totalrepro = r[10]
                cod_carrera = r[11]
                cod_mod = "ING01"
                ws.write(row_num, 0, u'%s' % cedula, formatocelda)
                ws.write(row_num, 1, u'%s' % estudiante, formatocelda)
                ws.write(row_num, 2, u'%s' % emailper, formatocelda)
                ws.write(row_num, 3, u'%s' % emailinst, formatocelda)
                ws.write(row_num, 4, u'%s' % telefono, formatocelda)
                ws.write(row_num, 5, u'%s' % carrera, formatocelda)
                ws.write(row_num, 6, u'%s' % modalidad, formatocelda)
                ws.write(row_num, 7, u'%s' % nivelmatricula, formatocelda)
                ws.write(row_num, 8, aniomalla, formatoceldaentero)
                ws.write(row_num, 9, totalaprobado, formatoceldaentero)
                ws.write(row_num, 10, totalrepro, formatoceldaentero)
                ws.write(row_num, 11, cod_carrera, formatoceldaentero)
                ws.write(row_num, 12, cod_mod, formatoceldaentero)
                ws.write(row_num, 13, id_matricula, formatoceldaentero)

                row_num += 1

            cursor.close()

            workbook.close()

            if notificacion_id > 0:
                eNotificacion = Notificacion.objects.get(pk=notificacion_id)
                eNotificacion.en_proceso = False
                eNotificacion.cuerpo = f'Generación del reporte de estudiantes con módulo de inglés aprobado del período académico {ePeriodo.nombre} listo'
                eNotificacion.url = "{}reportes/ingles/alumnos/{}".format(MEDIA_URL, nombre_archivo)
                eNotificacion.save(request)
            else:
                eNotificacion = Notificacion(cuerpo=f'Generación del reporte de estudiantes con módulo de inglés aprobado del período académico {ePeriodo.nombre} listo',
                                             titulo=f'Reporte de estudiantes con módulo de inglés aprobado del período académico {ePeriodo.nombre}',
                                             destinatario=ePersona,
                                             url="{}reportes/ingles/alumnos/{}".format(MEDIA_URL, nombre_archivo),
                                             prioridad=1,
                                             app_label='SGA',
                                             fecha_hora_visible=datetime.now() + timedelta(days=1),
                                             tipo=2,
                                             en_proceso=False)
                eNotificacion.save(request)
            try:
                send_user_notification(user=usernotify,
                                       payload={"head": f'Reporte de estudiantes con módulos de inglés aprobado',
                                                "body": f'Generación del reporte de estudiantes con módulos de inglés aprobado listo',
                                                "action": "notificacion",
                                                "timestamp": time.mktime(datetime.now().timetuple()),
                                                "url": "{}reportes/ingles/alumnos/{}".format(MEDIA_URL, nombre_archivo),
                                                "btn_notificaciones": traerNotificaciones(request, data, ePersona),
                                                "mensaje": 'Su reporte ha sido generado con exito'
                                                },
                                       ttl=500)
            except:
                pass
        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)
            if notificacion_id > 0:
                eNotificacion = Notificacion.objects.get(pk=notificacion_id)
                eNotificacion.titulo=f'Reporte de estudiantes con módulo de inglés aprobado del período académico {ePeriodo.nombre} falló en la ejecución'
                eNotificacion.en_proceso = False
                eNotificacion.cuerpo = textoerror
                eNotificacion.url = None
                eNotificacion.save(request)
            else:
                eNotificacion = Notificacion(cuerpo=textoerror,
                                             titulo=f'Reporte de estudiantes con módulo de inglés aprobado del período académico {ePeriodo.nombre} falló en la ejecución',
                                             destinatario=ePersona,
                                             url=None,
                                             prioridad=1,
                                             app_label='SGA',
                                             fecha_hora_visible=datetime.now() + timedelta(days=1),
                                             tipo=2,
                                             en_proceso=False)
                eNotificacion.save(request)
            try:
                send_user_notification(user=usernotify,
                                       payload={"head": f'Reporte de estudiantes con módulos de inglés aprobado fallido',
                                                "body": f'Generación del reporte de estudiantes con módulos de inglés aprobado a fallado',
                                                "action": "notificacion",
                                                "timestamp": time.mktime(datetime.now().timetuple()),
                                                "btn_notificaciones": traerNotificaciones(request, data, ePersona),
                                                "mensaje": textoerror,
                                                "error": True
                                                },
                                       ttl=500)
            except:
                pass


class reporte_distributivo_asignaturas(threading.Thread):

    def __init__(self, request, data, notiid, periodo):
        self.request = request
        self.data = data
        self.notiid = notiid
        self.periodo = periodo
        threading.Thread.__init__(self)

    def run(self):
        from sga.models import Persona, Notificacion
        try:
            directory = os.path.join(MEDIA_ROOT, 'reportes', 'matrices')
            request, data, notif, periodo = self.request, self.data, self.notiid, self.periodo
            usernotify = User.objects.get(pk=request.user.pk)
            pers = Persona.objects.get(usuario=usernotify)
            os.makedirs(directory, exist_ok=True)
            nombre_archivo = 'reporte_distributivo_asignaturas' + random.randint(1, 10000).__str__() + '.xlsx'
            directory = os.path.join(MEDIA_ROOT, 'reportes', 'matrices', nombre_archivo)
            __author__ = 'Unemi'
            workbook = xlsxwriter.Workbook(directory, {'constant_memory': True})
            ws = workbook.add_worksheet("Hoja1")

            fuentecabecera = workbook.add_format({
                'align': 'center',
                'bg_color': 'silver',
                'border': 1,
                'bold': 1
            })

            formatoceldacenter = workbook.add_format({
                'border': 1,
                'valign': 'vcenter',
                'align': 'center'})

            formatoceldacenter = workbook.add_format({
                'border': 1,
                'valign': 'vcenter',
                'align': 'center'})

            fuenteencabezado = workbook.add_format({
                'align': 'center',
                'bg_color': '#1C3247',
                'font_color': 'white',
                'border': 1,
                'font_size': 24,
                'bold': 1
            })

            columnas = [
                (u"FACULTAD", 60),
                (u"CARRERA", 60),
                (u"MODALIDAD", 60),
                (u"MALLA", 60),
                (u"SECCIÓN", 60),
                (u"NIVEL", 60),
                (u"PARALELO", 60),
                (u"ASIGNATURA", 60),
                (u"TEORICA PRACTICA", 60),
                (u"CUPO", 60),
                (u"MATRICULADOS", 60),
                (u"INSCRITOS", 60),  #
                (u"TOTAL MATRICULADOS", 60),
                (u"DOCENTE", 60),
                (u"CEDULA", 60),
                # (u"USUARIO", 6000),
                (u"AFINIDAD", 60),
                (u"HORAS SEMANALES", 40),
                (u"MALLA (HORAS PRESENCIALES SEMANALES)", 40),
                (u"TIPO", 40),
                (u"CORREO PERSONAL", 40),
                (u"CORREO INSTITUCIONAL", 40),
                (u"TIPO PROFESOR", 40),
                (u"PROFESOR DESDE", 40),
                (u"PROFESOR HASTA", 40),
                (u"DEDICACION", 50),
                (u"CATEGORIA", 40),
                (u"INICIO MATERIA", 40),
                (u"FIN MATERIA", 40),
                (u"FIN ASISTENCIA", 40),
                # (u"ID", 4000),
                (u"TELEFONO", 60),
                (u"MODELO EVALUATIVO", 60),
                (u"IDMATERIA", 25),
                (u"ACEPTACION", 25),
                (u"OBSERVACION ACEPTACION", 20),
                (u"HORARIO FECHA ACEPTACION", 20),
                (u"HORARIO ACEPTACION", 20),
                (u"HORARIO OBSERVACION ACEPTACION", 20),
                (u"IDMOODLE", 25),
                (u"PROVINCIA", 25),
                (u"CIUDAD", 25),
                (u"CANT ACEPTADOS", 60),
                (u"MODALIDAD IMPARTICION", 60),
                (u"ES PRÁCTICA", 40),
                (u"SILABO", 40),
                (u"TIENE GUIA PRÁCTICA", 40),
                (u"CERRADA", 40),
            ]
            ws.merge_range(0, 0, 0, columnas.__len__() - 1, 'UNIVERSIDAD ESTATAL ESTATAL DE MILAGRO', fuenteencabezado)
            ws.merge_range(1, 0, 1, columnas.__len__() - 1, f'DISTRIBUTIVO DE ASIGNATURAS',fuenteencabezado)
            row_num, numcolum = 2, 0
            for col_name in columnas:
                ws.write(row_num, numcolum, col_name[0], fuentecabecera)
                ws.set_column(numcolum, numcolum, col_name[1])
                numcolum += 1
            row_num += 1

            cursor = connections['default'].cursor()
            sql = f"""SET statement_timeout='20000 s';
                                        SELECT sga_coordinacion.nombre AS Facultad, 
                                            sga_carrera.nombre AS Carrera, sga_sesion.nombre AS Seccion, 
                                            sga_nivelmalla.nombre AS Nivel, sga_materia.paralelo AS Paralelo, 
                                            sga_materia.id AS Idmateria, sga_asignatura.nombre AS Asignatura, 
                                            sga_persona.apellido1 || ' ' || sga_persona.apellido2 || ' ' || sga_persona.nombres AS Docente, 
                                            sga_profesormateria.hora AS sga_profesormateria_hora, (CASE sga_profesormateria.principal WHEN TRUE THEN 'PRINCIPAL' ELSE 'PRACTICA' END) AS Tipo, sga_persona.cedula, (
                                        SELECT u.username
                                        FROM auth_user u
                                        WHERE u.id=sga_persona.usuario_id), sga_persona.email, sga_persona.emailinst, sga_materia.cupo AS cupo, (
                                        SELECT COUNT(*)
                                        FROM sga_materiaasignada ma, sga_matricula mat1
                                        WHERE ma.matricula_id=mat1.id AND ma.status=True AND mat1.status=True AND mat1.estado_matricula in (2,3) AND ma.materia_id=sga_materia.id AND ma.id NOT in (
                                        SELECT mr.materiaasignada_id
                                        FROM sga_materiaasignadaretiro mr)) AS nmatriculados, sga_tipoprofesor.nombre AS Tipoprofesor,
                                         sga_profesormateria.desde AS desde, sga_profesormateria.hasta AS hasta, (
                                        SELECT ti.nombre
                                        FROM sga_profesordistributivohoras dis,sga_tiempodedicaciondocente ti
                                        WHERE dis.dedicacion_id=ti.id AND dis.profesor_id=sga_profesor.id AND periodo_id={periodo} AND dis.status= TRUE) AS dedicacion, (
                                        SELECT ca.nombre
                                        FROM sga_profesordistributivohoras dis,sga_categorizaciondocente ca
                                        WHERE dis.categoria_id=ca.id AND dis.profesor_id=sga_profesor.id AND dis.periodo_id={periodo} AND dis.status= TRUE) AS categoria, 
                                        (CASE sga_asignaturamalla.practicas WHEN TRUE THEN 'SI' ELSE 'NO' END) AS tipomateria, 
                                        (CASE sga_profesormateria.afinidad WHEN TRUE THEN 'SI' ELSE 'NO' END) AS afinidad, 
                                        sga_materia.inicio AS inicio, sga_materia.fin AS fin, sga_materia.fechafinasistencias AS finasistencia, sga_materia.id AS id, sga_persona.telefono_conv AS telefonoconv, sga_persona.telefono AS telefono, EXTRACT(YEAR
                                        FROM sga_malla.inicio) AS anio, (
                                        SELECT modelo.nombre
                                        FROM sga_modeloevaluativo modelo
                                        WHERE modelo.id = sga_materia.modeloevaluativo_id) AS modeloevaluativo, sga_asignaturamalla.horaspresencialessemanales AS horaspresencialessemanales, sga_profesormateria.aceptarmateria AS aceptarmateria, sga_profesormateria.aceptarmateriaobs AS aceptarmateriaobs, sga_profesormateria.fecha_horario AS fecha_horario, sga_profesormateria.aceptarhorario AS aceptarhorario, sga_profesormateria.aceptarhorarioobs AS aceptarhorarioobs, sga_materia.idcursomoodle AS idcursomoodle, prov.nombre AS provincia, cant.nombre AS ciudad
                                        ,(
                                        SELECT COUNT(*)
                                        FROM sga_materiaasignada ma2, sga_matricula mat1
                                        WHERE ma2.matricula_id=mat1.id AND mat1.termino= TRUE AND ma2.materia_id=sga_materia.id AND ma2.id NOT in (
                                        SELECT mr.materiaasignada_id
                                        FROM sga_materiaasignadaretiro mr)) AS nmatriculados_acpta_termino,
                                        (
                                        SELECT COUNT(*)
                                        FROM sga_materiaasignada mas1, sga_matricula mat1, sga_nivel ni1
                                        WHERE mat1.estado_matricula=1 AND mas1.matricula_id=mat1.id AND mat1.nivel_id=ni1.id 
                                        AND ni1.periodo_id=sga_nivel.periodo_id AND mas1.materia_id=sga_materia.id) AS inscritos,
                                        (select CASE WHEN sga_detalleasignaturamallamodalidad.modalidad = 1 THEN 'PRESENCIAL' else 'VIRTUAL' END) as modalidad,
                                        (CASE sga_asignaturamalla.practicas 
                                        WHEN True THEN 'SI' ELSE 'NO' END) AS Malla_Practicas,
                                        (CASE 
                                            (SELECT count(*) 
                                                    FROM sga_silabo AS sga_s 
                                                    WHERE sga_s.materia_id = sga_materia.id AND sga_s.status=True AND sga_s.codigoqr=True
                                                ) 
                                            WHEN 0 THEN
                                                'NO'
                                            ELSE 
                                                'SI'
                                            END
                                        ) AS silabo,
                                        (SELECT 
                                            count(sga_tp.*) 
                                        FROM sga_tareapracticasilabosemanal AS sga_tp
                                        INNER JOIN sga_silabosemanal AS sga_ss ON sga_tp.silabosemanal_id = sga_ss.id
                                        INNER JOIN sga_silabo AS sga_s ON sga_ss.silabo_id = sga_s.id
                                        INNER JOIN sga_materia AS sga_m ON sga_m.id = sga_s.materia_id
                                        WHERE 
                                            sga_s.status= TRUE 
                                            AND sga_s.codigoqr= TRUE 
                                            AND sga_m.id = sga_materia.id 
                                            AND sga_tp.estado_id!=3 
                                            AND sga_tp.status=True
                                        ) AS trabajos_practicos,
                                        sga_materia.cerrado
                                        ,sga_carrera.modalidad
                                        FROM public.sga_materia sga_materia
                                        LEFT JOIN public.sga_profesormateria sga_profesormateria ON sga_materia.id = sga_profesormateria.materia_id AND sga_profesormateria.status= TRUE AND sga_profesormateria.activo= TRUE
                                        LEFT JOIN public.sga_profesor sga_profesor ON sga_profesor.id = sga_profesormateria.profesor_id AND sga_profesor.status= TRUE
                                        LEFT JOIN public.sga_tipoprofesor sga_tipoprofesor ON sga_tipoprofesor.id = sga_profesormateria.tipoprofesor_id AND sga_tipoprofesor.status= TRUE
                                        LEFT JOIN public.sga_persona sga_persona ON sga_profesor.persona_id = sga_persona.id AND sga_persona.status= TRUE
                                        LEFT JOIN sga_provincia prov ON prov.id=sga_persona.provincia_id
                                        LEFT JOIN sga_canton cant ON cant.id=sga_persona.canton_id
                                        INNER JOIN public.sga_nivel sga_nivel ON sga_materia.nivel_id = sga_nivel.id AND sga_nivel.status= TRUE
                                        INNER JOIN public.sga_asignatura sga_asignatura ON sga_materia.asignatura_id = sga_asignatura.id AND sga_asignatura.status= TRUE
                                        INNER JOIN public.sga_asignaturamalla sga_asignaturamalla ON sga_materia.asignaturamalla_id = sga_asignaturamalla.id AND sga_asignaturamalla.status= TRUE
                                        LEFT JOIN public.sga_detalleasignaturamallamodalidad sga_detalleasignaturamallamodalidad ON sga_asignaturamalla.id = sga_detalleasignaturamallamodalidad.asignaturamalla_id 
                                        INNER JOIN public.sga_nivelmalla sga_nivelmalla ON sga_asignaturamalla.nivelmalla_id = sga_nivelmalla.id AND sga_nivelmalla.status= TRUE
                                        INNER JOIN public.sga_malla sga_malla ON sga_asignaturamalla.malla_id = sga_malla.id AND sga_malla.status= TRUE
                                        INNER JOIN public.sga_carrera sga_carrera ON sga_malla.carrera_id = sga_carrera.id AND sga_carrera.status= TRUE
                                        INNER JOIN public.sga_coordinacion_carrera sga_coordinacion_carrera ON sga_carrera.id = sga_coordinacion_carrera.carrera_id
                                        INNER JOIN public.sga_coordinacion sga_coordinacion ON sga_coordinacion_carrera.coordinacion_id = sga_coordinacion.id AND sga_coordinacion.status= TRUE
                                        INNER JOIN public.sga_sesion sga_sesion ON sga_nivel.sesion_id = sga_sesion.id
                                        INNER JOIN public.sga_periodo sga_periodo ON sga_nivel.periodo_id = sga_periodo.id
                                        WHERE sga_periodo.id = {periodo} AND sga_materia.status= TRUE
                                        ORDER BY sga_coordinacion.nombre, sga_carrera.nombre, sga_sesion.nombre, sga_nivelmalla.nombre,sga_materia.paralelo,sga_asignatura.nombre
                                """
            cursor.execute(sql)
            results = cursor.fetchall()

            for r in results:
                campo1 = r[0].__str__()
                campo2 = r[1].__str__()
                campo3 = r[2].__str__()
                campo4 = r[3].__str__()
                campo5 = r[4].__str__()
                campo6 = r[5]
                campo7 = r[6].__str__()
                campo8 = r[7].__str__()
                campo9 = r[8]
                campo10 = r[9].__str__()
                campo11 = r[10].__str__()
                # campo12 = r[11]
                campo13 = r[12].__str__()
                campo14 = r[13].__str__()
                campo15 = int(r[14])
                campo16 = int(r[15])
                campo17 = r[16].__str__()
                campo18 = r[17].__str__()
                campo19 = r[18].__str__()
                campo20 = r[19].__str__()
                campo21 = r[20].__str__()
                campo22 = r[21].__str__()
                campo23 = r[22].__str__()
                campo24 = r[23].__str__()
                campo25 = r[24].__str__()
                campo26 = r[25].__str__()
                # campo27 = r[26]
                campo28 = r[27].__str__() + " - " + r[28].__str__()
                campo29 = r[29].__str__()
                campo30 = r[30].__str__()
                campo31 = r[31].__str__()
                if r[33] == None or r[33] == '':
                    campo32 = ''
                    campo33 = ''
                else:
                    campo32 = 'NO'
                    if r[32]:
                        campo32 = 'SI'
                    campo33 = r[33]
                campo34 = r[34]
                if r[36] == None or r[36] == '':
                    campo35 = ''
                    campo36 = ''
                else:
                    campo35 = 'NO'
                    if r[35]:
                        campo35 = 'SI'
                    campo36 = r[36]
                campo37 = r[37]
                campo38 = r[38]
                campo39 = r[39]
                campo_cant_acapta = r[40]
                inscritos = int(r[41])
                campo40 = r[42].__str__()
                campo41 = r[43]
                campo42 = r[44]
                campo43 = r[45]
                campo46 = r[46]
                campo47 = r[47]
                totalMatriculados = inscritos + int(campo16)
                modalidad = ""
                if campo47 == 1:
                    modalidad="PRESENCIAL"
                elif campo47 == 2:
                    modalidad = "SEMIPRESENCIAL"
                elif campo47 == 3:
                    modalidad = "EN LÍNEA"
                else:
                    modalidad = "HÍBRIDA"

                ws.write(row_num, 0, campo1, formatoceldacenter)
                ws.write(row_num, 1, campo2, formatoceldacenter)
                ws.write(row_num, 2, modalidad, formatoceldacenter)
                ws.write(row_num, 3, campo29, formatoceldacenter)
                ws.write(row_num, 4, campo3, formatoceldacenter)
                ws.write(row_num, 5, campo4, formatoceldacenter)
                ws.write(row_num, 6, campo5, formatoceldacenter)
                ws.write(row_num, 7, campo7, formatoceldacenter)
                ws.write(row_num, 8, campo22, formatoceldacenter)
                ws.write(row_num, 9, campo15, formatoceldacenter)
                ws.write(row_num, 10, campo16, formatoceldacenter)
                ws.write(row_num, 11, inscritos, formatoceldacenter)
                ws.write(row_num, 12, totalMatriculados, formatoceldacenter)

                ws.write(row_num, 13, campo8, formatoceldacenter)
                ws.write(row_num, 14, campo11, formatoceldacenter)
                # ws.write(row_num, 12, campo12, font_style2)
                ws.write(row_num, 15, campo23, formatoceldacenter)
                ws.write(row_num, 16, campo9, formatoceldacenter)
                ws.write(row_num, 17, campo9, formatoceldacenter)
                ws.write(row_num, 18, campo10, formatoceldacenter)
                ws.write(row_num, 19, campo13, formatoceldacenter)
                ws.write(row_num, 20, campo14, formatoceldacenter)
                ws.write(row_num, 21, campo17, formatoceldacenter)
                ws.write(row_num, 22, campo18, formatoceldacenter)
                ws.write(row_num, 23, campo19, formatoceldacenter)
                ws.write(row_num, 24, campo20, formatoceldacenter)
                ws.write(row_num, 25, campo21, formatoceldacenter)
                ws.write(row_num, 26, campo24, formatoceldacenter)
                ws.write(row_num, 27, campo25, formatoceldacenter)
                ws.write(row_num, 28, campo26, formatoceldacenter)
                # ws.write(row_num, 25, campo27, font_style2)
                ws.write(row_num, 29, campo28, formatoceldacenter)
                ws.write(row_num, 30, campo30, formatoceldacenter)
                ws.write(row_num, 31, campo6, formatoceldacenter)
                ws.write(row_num, 32, campo32, formatoceldacenter)
                ws.write(row_num, 33, campo33, formatoceldacenter)
                ws.write(row_num, 34, str(campo34), formatoceldacenter)
                ws.write(row_num, 35, campo35, formatoceldacenter)
                ws.write(row_num, 36, campo36, formatoceldacenter)
                ws.write(row_num, 37, campo37, formatoceldacenter)
                ws.write(row_num, 38, campo38, formatoceldacenter)
                ws.write(row_num, 39, campo39, formatoceldacenter)
                ws.write(row_num, 40, campo_cant_acapta, formatoceldacenter)
                ws.write(row_num, 41, campo40, formatoceldacenter)
                ws.write(row_num, 42, campo41, formatoceldacenter)
                ws.write(row_num, 43, campo42, formatoceldacenter)
                ws.write(row_num, 44, campo43, formatoceldacenter)
                ws.write(row_num, 45, campo46, formatoceldacenter)
                row_num += 1
            workbook.close()
            if notif > 0:
                noti = Notificacion.objects.get(pk=notif)
                noti.en_proceso = False
                noti.cuerpo = 'Excel Listo'
                noti.url = "{}reportes/matrices/{}".format(MEDIA_URL, nombre_archivo)
                noti.save()
            else:
                noti = Notificacion(cuerpo='Reporte Listo',
                                    titulo='Excel Distributivo de asignaturas',
                                    destinatario=pers, url="{}reportes/matrices/{}".format(MEDIA_URL, nombre_archivo),
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)
                noti.save(request)
            try:
                send_user_notification(user=usernotify, payload={
                    "head": "Reporte de distributivo de asignaturas",
                    "body": 'Excel terminado',
                    "action": "notificacion",
                    "timestamp": time.mktime(datetime.now().timetuple()),
                    "url": "{}reportes/matrices/{}".format(MEDIA_URL, nombre_archivo),
                    "btn_notificaciones": traerNotificaciones(request, data, pers),
                    "mensaje": 'Su reporte ha sido generado con exito'
                }, ttl=500)
            except:
                pass
        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)
            if notif > 0:
                noti = Notificacion.objects.get(pk=notif)
                noti.en_proceso = False
                noti.cuerpo = 'Error en el reporte {} - linea {}'.format(ex, sys.exc_info()[-1].tb_lineno)
                noti.url = ""
                noti.save()
            else:
                noti = Notificacion(cuerpo='Error en el reporte {} - linea {}'.format(ex, sys.exc_info()[-1].tb_lineno),
                                    titulo='Excel Distributivo de asignaturas',
                                    destinatario=pers, url="",
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)
                noti.save(request)
            try:
                send_user_notification(user=usernotify, payload={
                    "head": "Reporte de distributivo de asignaturas",
                    "body": 'Error en el reporte {} - linea {}'.format(ex, sys.exc_info()[-1].tb_lineno),
                    "action": "notificacion",
                    "timestamp": time.mktime(datetime.now().timetuple()),
                    "url": "",
                    "btn_notificaciones": traerNotificaciones(request, data, pers),
                    "mensaje": textoerror
                }, ttl=500)
            except:
                pass


class reporte_distributivo_preferencias(threading.Thread):

    def __init__(self, request, data, notiid, periodo):
        self.request = request
        self.data = data
        self.notiid = notiid
        self.periodo = periodo
        threading.Thread.__init__(self)

    def run(self):
        from sga.models import Persona, Notificacion
        try:
            directory = os.path.join(MEDIA_ROOT, 'reportes', 'matrices')
            request, data, notif, periodo = self.request, self.data, self.notiid, self.periodo
            usernotify = User.objects.get(pk=request.user.pk)
            pers = Persona.objects.get(usuario=usernotify)
            os.makedirs(directory, exist_ok=True)
            nombre_archivo = 'reporte_distributivo_preferencias' + random.randint(1, 10000).__str__() + '.xls'
            directory = os.path.join(MEDIA_ROOT, 'reportes', 'matrices', nombre_archivo)
            title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
            font_style = XFStyle()
            font_style.font.bold = True
            font_style2 = XFStyle()
            font_style2.font.bold = False
            wb = Workbook(encoding='utf-8')
            ws = wb.add_sheet('Hoja1')
            ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
            row_num = 3
            date_format = xlwt.XFStyle()
            date_format.num_format_str = 'yyyy/mm/dd'
            columns = [
                (u"FACULTAD", 6000), (u"CARRERA", 6000),
                (u"SECCIÓN", 6000), (u"NIVEL", 6000),
                (u"PARALELO", 6000), (u"IDMATERIA", 2500), (u"DOCENTE", 6000),
                (u"ASIGNATURA", 6000),
                (u"HORAS SEMANALES", 4000), (u"MALLA (HORAS PRESENCIALES SEMANALES)", 4000),
                (u"TIPO", 4000), (u"CEDULA", 4000),
                (u"TIPO PROFESOR", 4000), (u"DEDICACION", 5000),
                (u"CATEGORIA", 4000), (u"ACEPTACION", 2500),
                (u"OBSERVACION ACEPTACION", 10000), (u"HORARIO ACEPTACION", 10000),
                (u"HORARIO OBSERVACION ACEPTACION", 10000), (u"PREFERENCIA DE ASIGNATURA", 10000),
                (u"SI PREFERENCIA", 10000),
                (u"ASIGNATURA DE CONCURSO", 10000), (u"SI CONCURSO", 10000),
                (u"HISTORICO", 10000), (u"SI HISTORICO", 10000), (u"PERIODO", 10000)
            ]
            for col_num in range(len(columns)):
                ws.write(row_num, col_num, columns[col_num][0], font_style)
                ws.col(col_num).width = columns[col_num][1]
            cursor = connections['sga_select'].cursor()
            sql = """
                                SELECT sga_coordinacion.nombre AS Facultad, 
                                 sga_carrera.nombre AS Carrera, 
                                 sga_sesion.nombre AS Seccion, 
                                 sga_nivelmalla.nombre AS Nivel, 
                                 sga_materia.paralelo AS Paralelo, 
                                 sga_materia.id AS Idmateria, 
                                 sga_persona.apellido1 || ' ' || sga_persona.apellido2 || ' ' || sga_persona.nombres AS Docente, 
                                 sga_asignatura.nombre AS Asignatura, 
                                 sga_profesormateria.hora AS sga_profesormateria_hora, 
                                 sga_asignaturamalla.horaspresencialessemanales AS horaspresencialessemanales, 
                                 (CASE sga_profesormateria.principal WHEN TRUE THEN 'PRINCIPAL' ELSE 'PRACTICA' END) AS Tipo, 
                                 sga_persona.cedula,
                                 sga_tipoprofesor.nombre AS Tipoprofesor, 
                                (SELECT ti.nombre
                                FROM sga_profesordistributivohoras dis,sga_tiempodedicaciondocente ti
                                WHERE dis.dedicacion_id=ti.id AND dis.profesor_id=sga_profesor.id AND periodo_id= sga_periodo.id AND dis.status= TRUE LIMIT 1) AS dedicacion, 
                                (SELECT ca.nombre
                                FROM sga_profesordistributivohoras dis,sga_categorizaciondocente ca
                                WHERE dis.categoria_id=ca.id AND dis.profesor_id=sga_profesor.id AND dis.periodo_id=sga_periodo.id AND dis.status= TRUE LIMIT 1) AS categoria, 
                                 sga_profesormateria.aceptarmateria AS aceptarmateria, 
                                 sga_profesormateria.aceptarmateriaobs AS aceptarmateriaobs, 
                                 sga_profesormateria.aceptarhorario AS aceptarhorario, 
                                 sga_profesormateria.aceptarhorarioobs AS aceptarhorarioobs,
                                (SELECT string_agg(DISTINCT a.nombre, ', ')
                                FROM sga_asignaturamallapreferencia pre
                                INNER JOIN sga_asignaturamalla am ON am.id=pre.asignaturamalla_id
                                INNER JOIN sga_asignatura a ON a.id=am.asignatura_id
                                WHERE pre.periodo_id = sga_periodo.id AND pre.profesor_id=sga_profesor.id AND pre."status"= TRUE) AS preferencias,
                                (SELECT pre.id
                                FROM sga_asignaturamallapreferencia pre
                                INNER JOIN sga_asignaturamalla am ON am.id=pre.asignaturamalla_id
                                WHERE pre.periodo_id = sga_periodo.id AND pre.profesor_id=sga_profesor.id AND pre."status"= TRUE AND am.asignatura_id=sga_materia.asignatura_id LIMIT 1) AS sipreferencia,
                                (SELECT string_agg(DISTINCT a.nombre, ', ')
                                FROM sga_profesor_asignatura pa
                                INNER JOIN sga_asignatura a ON a.id=pa.asignatura_id
                                WHERE pa.profesor_id=sga_profesor.id) AS concurso,
                                (SELECT pa.id
                                FROM sga_profesor_asignatura pa
                                WHERE pa.profesor_id=sga_profesor.id AND pa.asignatura_id=sga_materia.asignatura_id LIMIT 1) AS concursosi,
                                (SELECT string_agg(DISTINCT a.nombre, ', ')
                                FROM sga_periodo per
                                INNER JOIN sga_nivel ni ON ni.periodo_id=per.id
                                INNER JOIN sga_materia ma ON ma.nivel_id=ni.id
                                INNER JOIN sga_asignaturamalla am ON am.id=ma.asignaturamalla_id
                                INNER JOIN sga_asignatura a ON a.id=am.asignatura_id
                                INNER JOIN sga_profesormateria pm ON pm.materia_id=ma.id
                                WHERE per.id = 90 AND pm.activo= TRUE AND pm.profesor_id=sga_profesor.id AND pm."status"= TRUE AND ma."status"= TRUE ) AS historico,
                                (SELECT ma.id
                                FROM sga_periodo per
                                INNER JOIN sga_nivel ni ON ni.periodo_id=per.id
                                INNER JOIN sga_materia ma ON ma.nivel_id=ni.id
                                INNER JOIN sga_profesormateria pm ON pm.materia_id=ma.id
                                WHERE per.id = 90 AND pm.activo= TRUE AND pm.profesor_id=sga_profesor.id AND ma.asignatura_id=sga_materia.asignatura_id AND pm."status"= TRUE AND ma."status"= TRUE
                                LIMIT 1 ) AS historicosi
                                FROM public.sga_materia sga_materia
                                LEFT JOIN public.sga_profesormateria sga_profesormateria ON sga_materia.id = sga_profesormateria.materia_id AND sga_profesormateria.status= TRUE AND sga_profesormateria.activo= TRUE
                                LEFT JOIN public.sga_profesor sga_profesor ON sga_profesor.id = sga_profesormateria.profesor_id AND sga_profesor.status= TRUE
                                LEFT JOIN public.sga_tipoprofesor sga_tipoprofesor ON sga_tipoprofesor.id = sga_profesormateria.tipoprofesor_id AND sga_tipoprofesor.status= TRUE
                                LEFT JOIN public.sga_persona sga_persona ON sga_profesor.persona_id = sga_persona.id AND sga_persona.status= TRUE
                                INNER JOIN public.sga_nivel sga_nivel ON sga_materia.nivel_id = sga_nivel.id AND sga_nivel.status= TRUE
                                INNER JOIN public.sga_asignatura sga_asignatura ON sga_materia.asignatura_id = sga_asignatura.id AND sga_asignatura.status= TRUE
                                INNER JOIN public.sga_asignaturamalla sga_asignaturamalla ON sga_materia.asignaturamalla_id = sga_asignaturamalla.id AND sga_asignaturamalla.status= TRUE
                                INNER JOIN public.sga_nivelmalla sga_nivelmalla ON sga_asignaturamalla.nivelmalla_id = sga_nivelmalla.id AND sga_nivelmalla.status= TRUE
                                INNER JOIN public.sga_malla sga_malla ON sga_asignaturamalla.malla_id = sga_malla.id AND sga_malla.status= TRUE
                                INNER JOIN public.sga_carrera sga_carrera ON sga_malla.carrera_id = sga_carrera.id AND sga_carrera.status= TRUE
                                INNER JOIN public.sga_coordinacion_carrera sga_coordinacion_carrera ON sga_carrera.id = sga_coordinacion_carrera.carrera_id
                                INNER JOIN public.sga_coordinacion sga_coordinacion ON sga_coordinacion_carrera.coordinacion_id = sga_coordinacion.id AND sga_coordinacion.status= TRUE
                                INNER JOIN public.sga_sesion sga_sesion ON sga_nivel.sesion_id = sga_sesion.id
                                INNER JOIN public.sga_periodo sga_periodo ON sga_nivel.periodo_id = sga_periodo.id AND sga_periodo.status= TRUE
                                WHERE sga_periodo.id = %s AND sga_materia.status= TRUE
                                ORDER BY sga_coordinacion.nombre, sga_carrera.nombre, sga_sesion.nombre, sga_nivelmalla.nombre,sga_materia.paralelo,sga_asignatura.nombre                    
                                """ % periodo
            cursor.execute(sql)
            results = cursor.fetchall()
            row_num = 4

            for r in results:
                i = 0
                campo1 = r[0].__str__()  # faculta
                campo2 = r[1].__str__()  # carrera
                campo3 = r[2].__str__()  # sesion
                campo4 = r[3].__str__()  # nivel
                campo5 = r[4].__str__()  # paralelo
                campo6 = r[5].__str__()  # salta numero
                campo7 = r[6].__str__()  # docente
                campo8 = r[7].__str__()  # asignatura
                campo9 = r[8]  # horas semanales
                campo10 = r[9]  # horas presenciales
                campo11 = r[10].__str__()  # tipo
                campo12 = r[11].__str__()  # cedula
                campo13 = r[12].__str__()  # TIPOPROFESOR
                campo14 = r[13].__str__()  # DEDICACION
                campo15 = r[14].__str__()  # CATEGORIA
                if r[16] == None or r[16] == '':
                    campo16 = ''
                elif r[15]:  # ACEPTACION
                    campo16 = 'SI'
                else:
                    campo16 = 'NO'
                campo17 = r[16].__str__() if r[16] else ''
                if r[18] == None or r[18] == '':
                    campo18 = ''
                elif r[17]:  # ACEPTACION HORARIO
                    campo18 = 'SI'
                else:
                    campo18 = 'NO'
                campo19 = r[18].__str__() if r[18] else ''
                campo20 = r[19].__str__() if r[19] else ''
                if r[20] != None:  # SI PREFERENCIAS
                    campo21 = 'SI'
                else:
                    campo21 = 'NO'

                campo22 = r[21].__str__() if r[21] else ''
                if r[22] != None:  # SI CONCURSO
                    campo23 = 'SI'
                else:
                    campo23 = 'NO'

                campo24 = r[23].__str__() if r[23] else ''
                if r[24] != None:  # SI HISTORICO
                    campo25 = 'SI'
                else:
                    campo25 = 'NO'
                campo26 = '%s' % periodo

                ws.write(row_num, 0, campo1, font_style2)
                ws.write(row_num, 1, campo2, font_style2)
                ws.write(row_num, 2, campo3, font_style2)
                ws.write(row_num, 3, campo4, font_style2)
                ws.write(row_num, 4, campo5, font_style2)
                ws.write(row_num, 5, campo6, font_style2)
                ws.write(row_num, 6, campo7, font_style2)
                ws.write(row_num, 7, campo8, font_style2)
                ws.write(row_num, 8, campo9, font_style2)
                ws.write(row_num, 9, campo10, font_style2)
                ws.write(row_num, 10, campo11, font_style2)
                ws.write(row_num, 11, campo12, font_style2)
                ws.write(row_num, 12, campo13, font_style2)
                ws.write(row_num, 13, campo14, font_style2)
                ws.write(row_num, 14, campo15, font_style2)
                ws.write(row_num, 15, campo16, font_style2)
                ws.write(row_num, 16, campo17, font_style2)
                ws.write(row_num, 17, campo18, font_style2)
                ws.write(row_num, 18, campo19, font_style2)
                ws.write(row_num, 19, campo20, font_style2)
                ws.write(row_num, 20, campo21, font_style2)
                ws.write(row_num, 21, campo22, font_style2)
                ws.write(row_num, 22, campo23, font_style2)
                ws.write(row_num, 23, campo24, font_style2)
                ws.write(row_num, 24, campo25, font_style2)
                ws.write(row_num, 25, campo26, font_style2)
                row_num += 1
            wb.save(directory)
            if notif > 0:
                noti = Notificacion.objects.get(pk=notif)
                noti.en_proceso = False
                noti.cuerpo = 'Excel Listo'
                noti.url = "{}reportes/matrices/{}".format(MEDIA_URL, nombre_archivo)
                noti.save()
            else:
                noti = Notificacion(cuerpo='Reporte Listo',
                                    titulo='Excel Distributivo de asignaturas de preferencias',
                                    destinatario=pers, url="{}reportes/matrices/{}".format(MEDIA_URL, nombre_archivo),
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)
                noti.save(request)
            try:
                send_user_notification(user=usernotify, payload={
                    "head": noti.titulo,
                    "body": noti.cuerpo,
                    "action": "notificacion",
                    "timestamp": time.mktime(datetime.now().timetuple()),
                    "url": noti.url,
                    "btn_notificaciones": traerNotificaciones(request, data, pers),
                    "mensaje": 'Su reporte ha sido generado con exito'
                }, ttl=500)
            except:
                pass
        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)
            if notif > 0:
                noti = Notificacion.objects.get(pk=notif)
                noti.en_proceso = False
                noti.cuerpo = 'Error en el reporte {} - linea {}'.format(ex, sys.exc_info()[-1].tb_lineno)
                noti.url = ""
                noti.save()
            else:
                noti = Notificacion(cuerpo='Error en el reporte {} - linea {}'.format(ex, sys.exc_info()[-1].tb_lineno),
                                    titulo='Excel Distributivo de asignaturas de preferencias',
                                    destinatario=pers, url="",
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)
                noti.save(request)
            try:
                send_user_notification(user=usernotify, payload={
                    "head": noti.titulo,
                    "body": noti.cuerpo,
                    "action": "notificacion",
                    "timestamp": time.mktime(datetime.now().timetuple()),
                    "url": noti.url,
                    "btn_notificaciones": traerNotificaciones(request, data, pers),
                    "mensaje": textoerror
                }, ttl=500)
            except:
                pass


class reporte_horarios_y_aulas(threading.Thread):

    def __init__(self, request, data, notiid, periodo):
        self.request = request
        self.data = data
        self.periodo = periodo
        self.notiid = notiid
        threading.Thread.__init__(self)

    def run(self):
        from sga.models import TIPOHORARIO, Persona, Notificacion
        try:
            directory = os.path.join(MEDIA_ROOT, 'reportes', 'horarios')
            request, data, notiid, periodo = self.request, self.data, self.notiid, self.periodo
            os.makedirs(directory, exist_ok=True)
            nombre_archivo = "REPORTE_HORARIOS_Y_AULAS{}.xls".format(random.randint(1, 10000).__str__())
            directory = os.path.join(MEDIA_ROOT, 'reportes', 'horarios', nombre_archivo)
            style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
            style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
            style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
            title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
            style1 = easyxf(num_format_str='D-MMM-YY')
            font_style = XFStyle()
            font_style.font.bold = True
            font_style2 = XFStyle()
            font_style2.font.bold = False

            wb = Workbook(encoding='utf-8')
            ws = wb.add_sheet("datos")

            ws.write_merge(0, 0, 0, 12, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
            columns = [
                (u"AULA", 6000),
                (u"CAPACIDAD AULA", 6000),
                (u"CUPO MATRICULA", 6000),
                (u"MATRICULADOS", 6000),
                (u"INSCRITOS", 6000),
                (u"DÍAS", 6000),
                (u"INICIO HORARIO", 6000),
                (u"FIN HORARIO", 6000),
                (u"HORARIOS", 6000),
                (u"ASIGNATURA", 6000),
                (u"MODALIDAD", 6000),
                (u"IDMATERIA", 6000),
                (u"JORNADA", 6000),
                (u"NIVEL", 6000),
                (u"CARRERA", 6000),
                (u"PARALELO", 6000),
                (u"DOCENTE", 6000),
                (u"TIPO PROFESOR", 6000),
                (u"CATEGORIA", 6000),
                (u"DEDICACIÓN", 6000),
                (u"PRINCIPAL", 6000),
                (u"FACULTAD", 6000),
                (u"TIPO MATERIA", 6000),
                (u"ID HORARIO", 4000),
                (u"INICIO MATERIA", 4000),
                (u"FIN MATERIA", 4000),
                (u"TIPO HORARIO", 4000),
                (u"ES PRÁCTICA", 4000),
                # (u"SILABO", 4000),
                # (u"TIENE GUIA PRÁCTICA", 4000),
                # (u"ID MATERIA", 4000),

            ]
            row_num = 3
            for col_num in range(len(columns)):
                ws.write(row_num, col_num, columns[col_num][0], font_style)
                ws.col(col_num).width = columns[col_num][1]
            cursor = connections['sga_select'].cursor()
            # lista_json = []
            # data = {}

            sql = f""" 
                                    SET statement_timeout = '36000s';
                                    SELECT al.nombre AS Aula, al.capacidad AS capacidad_aula, mat.cupo AS cupo_matriculas, 
                                     (
                                    SELECT COUNT(*)
                                    FROM sga_materiaasignada mas1, sga_matricula mat1, sga_nivel ni1
                                    WHERE mat1.estado_matricula in (2,3) AND mas1.matricula_id=mat1.id AND mat1.nivel_id=ni1.id AND ni1.periodo_id=ni.periodo_id AND mas1.materia_id=mat.id) AS Matriculados,
                                     (
                                        SELECT COUNT(*)
                                        FROM sga_materiaasignada mas1, sga_matricula mat1, sga_nivel ni1
                                        WHERE mat1.estado_matricula=1 AND mas1.matricula_id=mat1.id AND mat1.nivel_id=ni1.id 
                                        AND ni1.periodo_id=ni.periodo_id AND mas1.materia_id=mat.id) AS inscritos, 
                                     (CASE cl.dia WHEN 1 THEN 'LUNES' WHEN 2 THEN 'MARTES' WHEN 3 THEN 'MIERCOLES' WHEN 4 THEN 'JUEVES' WHEN 5 THEN 'VIERNES' WHEN 6 THEN 'SABADO' WHEN 7 THEN 'DOMINGO' END) AS dia, 
                                     cl.inicio, cl.fin, (tu.comienza|| '  ' || tu.termina) AS Horario, asi.nombre AS Asignatura, 
                                     (select CASE WHEN mod.modalidad = 1 THEN 'PRESENCIAL' else 'VIRTUAL' END) as modalidad,
                                     mat.id AS idmateria,
                                     ni.paralelo AS jornada, 
                                     nmall.nombre AS nivel, 
                                     ca.nombre AS Carrera, mat.paralelo AS paralelo, per.apellido1 ||' '|| per.apellido2 ||' '|| per.nombres AS docente,
                                     tipop.nombre AS tipo_profesor, cat.nombre AS categorizacion,ded.nombre AS dedicacion,
                                     (CASE pm.principal WHEN TRUE THEN 'SI' ELSE 'NO' END) AS principal, 
                                     cor.nombre AS facultad, 
                                     (CASE asimall.practicas WHEN TRUE THEN 'SI' ELSE 'NO' END) AS tipomateria, 
                                     cl.id AS Id, 
                                     mat.inicio AS Inicio_materia, 
                                     mat.fin AS Fin_materia, 
                                     mat.id AS id_materia,
                                     cl.tipohorario AS cl_tipohorario,
                                     asimall.practicas AS asimall_practicas
                                     /*(CASE 
                                        (SELECT count(*) 
                                                FROM sga_silabo AS sga_s 
                                                WHERE sga_s.materia_id = pm.materia_id AND sga_s.status=True AND sga_s.codigoqr=True
                                            ) 
                                        WHEN 0 THEN
                                            'NO'
                                        ELSE 
                                            'SI'
                                        END
                                    ) AS silabo,
                                    (SELECT 
                                        count(sga_tp.*) 
                                    FROM sga_tareapracticasilabosemanal AS sga_tp
                                    INNER JOIN sga_silabosemanal AS sga_ss ON sga_tp.silabosemanal_id = sga_ss.id
                                    INNER JOIN sga_silabo AS sga_s ON sga_ss.silabo_id = sga_s.id
                                    INNER JOIN sga_materia AS sga_m ON sga_m.id = sga_s.materia_id
                                    WHERE 
                                        sga_s.status= TRUE 
                                        AND sga_s.codigoqr= TRUE 
                                        AND sga_m.id = mat.id 
                                        AND sga_tp.estado_id!=3 
                                        AND sga_tp.status=True
                                    ) AS trabajos_practicos*/
                                    FROM sga_clase cl 
            						INNER JOIN sga_aula al ON cl.aula_id=al.id
            						INNER JOIN sga_materia mat ON cl.materia_id=mat.id
            						INNER JOIN sga_nivel ni ON mat.nivel_id=ni.id
            						INNER JOIN sga_asignatura asi ON mat.asignatura_id=asi.id
            						INNER JOIN sga_asignaturamalla asimall ON mat.asignaturamalla_id = asimall.id
            						INNER JOIN sga_nivelmalla nmall ON asimall.nivelmalla_id = nmall.id 
                                    INNER JOIN sga_malla mall ON asimall.malla_id = mall.id 
                                    INNER JOIN sga_carrera ca ON mall.carrera_id=ca.id
                                    INNER JOIN sga_turno tu ON tu.id=cl.turno_id 
                                    INNER JOIN sga_profesormateria pm ON pm.materia_id=mat.id
                                    INNER JOIN sga_tipoprofesor tipop ON tipop.id=pm.tipoprofesor_id
                                    INNER JOIN sga_profesor pr ON pr.id=pm.profesor_id AND cl.profesor_id=pr.id
                                    INNER JOIN sga_tiempodedicaciondocente ded ON pr.dedicacion_id=ded.id
                                    INNER JOIN sga_categorizaciondocente cat ON pr.categoria_id=cat.id
                                    INNER JOIN sga_persona per ON per.id=pr.persona_id
                                    INNER JOIN sga_coordinacion_carrera cc ON cc.carrera_id=ca.id
                           			INNER JOIN sga_coordinacion cor ON cor.id=cc.coordinacion_id 
                           			LEFT JOIN sga_detalleasignaturamallamodalidad mod ON asimall.id = mod.asignaturamalla_id 
                                    WHERE cl.activo= TRUE AND ni.periodo_id={periodo}
                                """
            cursor.execute(sql)
            results = cursor.fetchall()
            row_num = 4

            for r in results:
                i = 0
                campo1 = r[0]
                campo2 = r[1]
                campo3 = r[2]
                campo4 = r[3]
                campo5 = r[4]
                campo6 = r[5]
                campo7 = r[6]
                campo8 = r[7]
                campo9 = r[8]
                campo10 = r[9]
                campo11 = r[10]
                campo12 = r[11]
                campo13 = r[12]
                campo14 = r[13]
                campo15 = r[14]
                campo16 = r[15]
                campo17 = r[16]
                campo18 = r[17]
                campo19 = r[18]
                campo20 = r[19]
                campo21 = r[20]
                campo22 = r[21]
                campo23 = r[22]
                campo24 = r[23]
                campo25 = r[24]
                campo26 = r[25]
                campo27 = dict(TIPOHORARIO)[r[27]]
                campo28 = 'SI' if r[28] else 'NO'
                # campo29 = r[29]
                # campo30 = r[30]
                ws.write(row_num, 0, u"%s" % campo1, font_style2)
                ws.write(row_num, 1, u"%s" % campo2, font_style2)
                ws.write(row_num, 2, u"%s" % campo3, font_style2)
                ws.write(row_num, 3, u"%s" % campo4, font_style2)
                ws.write(row_num, 4, u"%s" % campo5, font_style2)
                ws.write(row_num, 5, u"%s" % campo6, style1)
                ws.write(row_num, 6, u"%s" % campo7, style1)
                ws.write(row_num, 7, u"%s" % campo8, font_style2)
                ws.write(row_num, 8, u"%s" % campo9, font_style2)
                ws.write(row_num, 9, u"%s" % campo10, font_style2)
                ws.write(row_num, 10, u"%s" % campo11, font_style2)
                ws.write(row_num, 11, u"%s" % campo12, font_style2)
                ws.write(row_num, 12, u"%s" % campo13, font_style2)
                ws.write(row_num, 13, u"%s" % campo14, font_style2)
                ws.write(row_num, 14, u"%s" % campo15, font_style2)
                ws.write(row_num, 15, u"%s" % campo16, font_style2)
                ws.write(row_num, 16, u"%s" % campo17, font_style2)
                ws.write(row_num, 17, u"%s" % campo18, font_style2)
                ws.write(row_num, 18, u"%s" % campo19, style1)
                ws.write(row_num, 19, u"%s" % campo20, style1)
                ws.write(row_num, 20, u"%s" % campo21, font_style2)
                ws.write(row_num, 21, u"%s" % campo22, font_style2)
                ws.write(row_num, 22, u"%s" % campo23, font_style2)
                ws.write(row_num, 23, u"%s" % campo24, font_style2)
                ws.write(row_num, 24, u"%s" % campo25, font_style2)
                ws.write(row_num, 25, u"%s" % campo26, font_style2)
                ws.write(row_num, 26, u"%s" % campo27, font_style2)
                ws.write(row_num, 27, u"%s" % campo28, font_style2)
                # ws.write(row_num, 28, u"%s" % campo29, font_style2)
                # ws.write(row_num, 29, u"%s" % campo30, font_style2)
                row_num += 1

            wb.save(directory)
            usernotify = User.objects.get(pk=request.user.pk)
            pers = Persona.objects.get(usuario=usernotify)
            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.cuerpo = 'Reporte Listo'
                noti.url = "{}reportes/horarios/{}".format(MEDIA_URL, nombre_archivo)
                noti.save()
            else:
                noti = Notificacion(cuerpo='Reporte Listo', titulo='Reporte de horarios y aulas',
                                    destinatario=pers, url="{}reportes/horarios/{}".format(MEDIA_URL, nombre_archivo),
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)
                noti.save(request)
            try:
                send_user_notification(user=usernotify, payload={
                    "head": noti.titulo,
                    "body": noti.cuerpo,
                    "action": "notificacion",
                    "timestamp": time.mktime(datetime.now().timetuple()),
                    "url": noti.url,
                    "btn_notificaciones": traerNotificaciones(request, data, pers),
                    "mensaje": 'Su reporte ha sido generado con exito'
                }, ttl=500)
            except:
                pass
        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)
            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.cuerpo = 'Error en el reporte {} - linea {}'.format(ex, sys.exc_info()[-1].tb_lineno)
                noti.url = ""
                noti.save()
            else:
                noti = Notificacion(cuerpo='Error en el reporte {} - linea {}'.format(ex, sys.exc_info()[-1].tb_lineno),
                                    titulo='Reporte de horarios y aulas',
                                    destinatario=pers, url="",
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)
                noti.save(request)
            try:
                send_user_notification(user=usernotify, payload={
                    "head": noti.titulo,
                    "body": noti.cuerpo,
                    "action": "notificacion",
                    "timestamp": time.mktime(datetime.now().timetuple()),
                    "url": noti.url,
                    "btn_notificaciones": traerNotificaciones(request, data, pers),
                    "mensaje": textoerror
                }, ttl=500)
            except:
                pass


class reporte_horarios_aulas_y_actividad(threading.Thread):

    def __init__(self, request, data, notiid, periodo):
        self.request = request
        self.periodo = periodo
        self.notiid = notiid
        threading.Thread.__init__(self)

    def run(self):
        from sga.models import TIPOHORARIO, Persona, Notificacion
        try:
            directory = os.path.join(MEDIA_ROOT, 'reportes', 'horarios')
            request, data, notiid, periodo = self.request, self.data, self.notiid, self.periodo
            os.makedirs(directory, exist_ok=True)
            nombre_archivo = "REPORTE_HORARIOS_AULAS_Y_ACTIVIDADES_{}.xls".format(random.randint(1, 10000).__str__())
            directory = os.path.join(MEDIA_ROOT, 'reportes', 'horarios', nombre_archivo)
            style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
            style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
            style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
            title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
            style1 = easyxf(num_format_str='D-MMM-YY')
            font_style = XFStyle()
            font_style.font.bold = True
            font_style2 = XFStyle()
            font_style2.font.bold = False
            wb = Workbook(encoding='utf-8')
            ws = wb.add_sheet("datos")
            ws.write_merge(0, 0, 0, 12, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
            columns = [
                (u"TIPO", 3000),
                (u"AULA", 6000),
                (u"CAPACIDAD AULA", 6000),
                (u"CUPO MATRICULA", 6000),
                (u"MATRICULADOS", 6000),
                (u"DÍAS", 6000),
                (u"INICIO HORARIO", 6000),
                (u"FIN HORARIO", 6000),
                (u"HORARIOS", 6000),
                (u"ASIGNATURA/ACTIVIDAD", 6000),
                (u"TEORICA PRACTICA", 6000),
                (u"JORNADA", 6000),
                (u"NIVEL", 6000),
                (u"PARALELO", 6000),
                (u"CARRERA", 6000),
                (u"DOCENTE", 6000),
                (u"PRINCIPAL", 6000),
                (u"FACULTAD", 6000),
                # (u"TIPO MATERIA", 6000),
                (u"ID HORARIO", 4000),
                (u"INICIO MATERIA", 4000),
                (u"FIN MATERIA", 4000),
                # (u"ID MATERIA", 4000),
                (u"ESTADO SOLICITUD ACTIVIDAD", 6000)

            ]
            row_num = 3
            for col_num in range(len(columns)):
                ws.write(row_num, col_num, columns[col_num][0], font_style)
                ws.col(col_num).width = columns[col_num][1]
            cursor = connections['sga_select'].cursor()
            # lista_json = []
            # data = {}
            sql = "select 'Materias', al.nombre as Aula, al.capacidad as capacidad_aula, mat.cupo as cupo_matriculas," \
                  " (select count(*) from sga_materiaasignada mas1, sga_matricula mat1, sga_nivel ni1 where mat1.estado_matricula in (2,3) and mas1.matricula_id=mat1.id and mat1.nivel_id=ni1.id and ni1.periodo_id=ni.periodo_id and mas1.materia_id=mat.id) as Matriculados," \
                  " (case cl.dia  when 1 then 'LUNES' when 2 then 'MARTES' when 3 then 'MIERCOLES' when 4 then 'JUEVES' when 5 then 'VIERNES' when 6 then 'SABADO' when 7 then 'DOMINGO'  end) as dia ," \
                  " cl.inicio , cl.fin , (tu.comienza|| '  ' || tu.termina) as Horario, asi.nombre as Asignatura , ni.paralelo as jornada , nmall.nombre as nivel , ca.nombre as Carrera, mat.paralelo as paralelo , per.apellido1 ||' '|| per.apellido2 ||' '|| per.nombres as docente," \
                  " (case pm.principal when true then 'SI' else 'NO' end) as principal, cor.nombre as facultad, (case asimall.practicas when true then 'SI' else 'NO' end) as tipomateria, cl.id as Id, mat.inicio as Inicio_materia, mat.fin as Fin_materia, mat.id as id_materia, null " \
                  " from sga_clase cl, sga_aula al, sga_materia mat, sga_nivel ni, sga_asignatura asi ,sga_asignaturamalla asimall, sga_nivelmalla nmall, sga_malla mall , sga_carrera ca ,  sga_turno tu , sga_profesormateria pm," \
                  " sga_profesor pr, sga_persona per, sga_coordinacion_carrera cc, sga_coordinacion cor " \
                  " where cl.aula_id=al.id and cl.materia_id=mat.id and mat.nivel_id=ni.id and mat.asignatura_id=asi.id and cl.activo=true and  mat.asignaturamalla_id = asimall.id and" \
                  " pm.materia_id=mat.id and pr.id=pm.profesor_id and per.id=pr.persona_id and asimall.nivelmalla_id = nmall.id and asimall.malla_id = mall.id and tu.id=cl.turno_id and  mall.carrera_id=ca.id" \
                  " and cc.carrera_id=ca.id  and cor.id=cc.coordinacion_id and ni.periodo_id='" + str(periodo) + "'" \
                                                                                                            "union " \
                                                                                                            "select 'Docencia', null as Aula, null as capacidad_aula, null as cupo_matriculas, null as Matriculados,  " \
                                                                                                            "(case actividad.dia  when 1 then 'LUNES' when 2 then 'MARTES' when 3 then 'MIERCOLES' when 4 then 'JUEVES' when 5 then 'VIERNES' when 6 then 'SABADO' when 7 then 'DOMINGO'  end) as dia ,  " \
                                                                                                            "actividad.inicio, actividad.fin, (tur.comienza|| '  ' || tur.termina) as Horario,  " \
                                                                                                            "critdoc.nombre as Asignatura, null as jornada, null as nivel, null as Carrera,  " \
                                                                                                            "null as paralelo , per.apellido1 ||' '|| per.apellido2 ||' '|| per.nombres as docente, null as principal, null as facultad,  " \
                                                                                                            "null as tipomateria, null as Id, null as Inicio_materia, null as Fin_materia, null as id_materia, (case actividad.estadosolicitud  when 1 then 'SOLICITADO' when 2 then 'APROBADO' when 3 then 'RECHAZADO'  end) as estadosolicitud  " \
                                                                                                            "from 	sga_claseactividad actividad  " \
                                                                                                            "inner join sga_detalledistributivo dist on actividad.detalledistributivo_id = dist.id " \
                                                                                                            "inner join sga_criteriodocenciaperiodo critdocper on dist.criteriodocenciaperiodo_id = critdocper.id " \
                                                                                                            "inner join sga_criteriodocencia critdoc on critdocper.criterio_id = critdoc.id " \
                                                                                                            "inner join sga_profesordistributivohoras pdh on dist.distributivo_id = pdh.id " \
                                                                                                            "inner join sga_profesor pf on pdh.profesor_id = pf.id  " \
                                                                                                            "inner join sga_persona per on pf.persona_id = per.id " \
                                                                                                            "inner join sga_turno tur on actividad.turno_id = tur.id  " \
                                                                                                            "where pdh.periodo_id = '" + str(periodo) + "' and actividad.status=True and actividad.activo=True	and dist.criteriogestionperiodo_id is null and dist.criterioinvestigacionperiodo_id is null " \
                                                                                                                                                   "union " \
                                                                                                                                                   "select 'Gestion', null as Aula, 	null as capacidad_aula, null as cupo_matriculas, null as Matriculados,  " \
                                                                                                                                                   "(case actividad.dia  when 1 then 'LUNES' when 2 then 'MARTES' when 3 then 'MIERCOLES' when 4 then 'JUEVES' when 5 then 'VIERNES' when 6 then 'SABADO' when 7 then 'DOMINGO'  end) as dia ,  " \
                                                                                                                                                   "actividad.inicio, actividad.fin, (tur.comienza|| '  ' || tur.termina) as Horario, critdoc.nombre as Asignatura ,  " \
                                                                                                                                                   "null as jornada , null as nivel , null as Carrera, null as paralelo ,  " \
                                                                                                                                                   "per.apellido1 ||' '|| per.apellido2 ||' '|| per.nombres as docente, null as principal, null as facultad,  " \
                                                                                                                                                   "null as tipomateria, null as Id, 	null as Inicio_materia, 	null as Fin_materia, null as id_materia, (case actividad.estadosolicitud  when 1 then 'SOLICITADO' when 2 then 'APROBADO' when 3 then 'RECHAZADO'  end) as estadosolicitud    " \
                                                                                                                                                   "from 	sga_claseactividad actividad " \
                                                                                                                                                   "inner join sga_detalledistributivo dist on actividad.detalledistributivo_id = dist.id " \
                                                                                                                                                   "inner join sga_criteriogestionperiodo critdocper on dist.criteriogestionperiodo_id = critdocper.id " \
                                                                                                                                                   "inner join sga_criteriogestion critdoc on critdocper.criterio_id = critdoc.id " \
                                                                                                                                                   "inner join sga_profesordistributivohoras pdh on dist.distributivo_id = pdh.id " \
                                                                                                                                                   "inner join sga_profesor pf on pdh.profesor_id = pf.id  " \
                                                                                                                                                   "inner join sga_persona per on pf.persona_id = per.id " \
                                                                                                                                                   "inner join sga_turno tur on actividad.turno_id = tur.id  " \
                                                                                                                                                   "where pdh.periodo_id = '" + str(periodo) + "' and actividad.status=True and actividad.activo=True	and dist.criteriodocenciaperiodo_id is null and dist.criterioinvestigacionperiodo_id is null " \
                                                                                                                                                                                          "union " \
                                                                                                                                                                                          "select 'Investigación', null as Aula, 	null as capacidad_aula, 	null as cupo_matriculas,  null as Matriculados,  " \
                                                                                                                                                                                          "(case actividad.dia  when 1 then 'LUNES' when 2 then 'MARTES' when 3 then 'MIERCOLES' when 4 then 'JUEVES' when 5 then 'VIERNES' when 6 then 'SABADO' when 7 then 'DOMINGO'  end) as dia ,  " \
                                                                                                                                                                                          "actividad.inicio , 	actividad.fin , 	(tur.comienza|| '  ' || tur.termina) as Horario, critdoc.nombre as Asignatura ,  " \
                                                                                                                                                                                          "null as jornada , null as nivel , null as Carrera, null as paralelo ,  " \
                                                                                                                                                                                          "per.apellido1 ||' '|| per.apellido2 ||' '|| per.nombres as docente,  null as principal, null as facultad, 	null as tipomateria, 	null as Id, 	null as Inicio_materia, 	null as Fin_materia,	null as id_materia , (case actividad.estadosolicitud  when 1 then 'SOLICITADO' when 2 then 'APROBADO' when 3 then 'RECHAZADO'  end) as estadosolicitud   " \
                                                                                                                                                                                          "from	sga_claseactividad actividad " \
                                                                                                                                                                                          "inner join sga_detalledistributivo dist on actividad.detalledistributivo_id = dist.id " \
                                                                                                                                                                                          "inner join sga_criterioinvestigacionperiodo critdocper on dist.criterioinvestigacionperiodo_id = critdocper.id " \
                                                                                                                                                                                          "inner join sga_criterioinvestigacion critdoc on critdocper.criterio_id = critdoc.id " \
                                                                                                                                                                                          "inner join sga_profesordistributivohoras pdh on dist.distributivo_id = pdh.id " \
                                                                                                                                                                                          "inner join sga_profesor pf on pdh.profesor_id = pf.id  " \
                                                                                                                                                                                          "inner join sga_persona per on pf.persona_id = per.id " \
                                                                                                                                                                                          "inner join sga_turno tur on actividad.turno_id = tur.id  " \
                                                                                                                                                                                          "where pdh.periodo_id = '" + str(periodo) + "' and actividad.status=True and actividad.activo=True	and dist.criteriodocenciaperiodo_id is null and dist.criteriogestionperiodo_id is null "
            cursor.execute(sql)
            results = cursor.fetchall()
            row_num = 4
            for r in results:
                i = 0
                campo1 = r[1]
                campo2 = r[2]
                campo3 = r[3]
                campo4 = r[4]
                campo5 = r[5]
                campo6 = r[6]
                campo7 = r[7]
                campo8 = r[8]
                campo9 = r[9]
                campo10 = r[10]
                campo11 = r[11]
                campo12 = r[13]
                campo13 = r[12]
                campo14 = r[14]
                campo15 = r[15]
                campo16 = r[16]
                campo17 = r[17]
                campo18 = r[18]
                campo19 = r[19]
                campo20 = r[20]
                campo21 = r[0]
                campo22 = r[22]

                ws.write(row_num, 0, campo21, font_style2)
                ws.write(row_num, 1, campo1, font_style2)
                ws.write(row_num, 2, campo2, font_style2)
                ws.write(row_num, 3, campo3, font_style2)
                ws.write(row_num, 4, campo4, font_style2)
                ws.write(row_num, 5, campo5, font_style2)
                ws.write(row_num, 6, campo6, style1)
                ws.write(row_num, 7, campo7, style1)
                ws.write(row_num, 8, campo8, font_style2)
                ws.write(row_num, 9, campo9, font_style2)
                ws.write(row_num, 10, campo17, font_style2)
                ws.write(row_num, 11, campo10, font_style2)
                ws.write(row_num, 12, campo11, font_style2)
                ws.write(row_num, 13, campo12, font_style2)
                ws.write(row_num, 14, campo13, font_style2)
                ws.write(row_num, 15, campo14, font_style2)
                ws.write(row_num, 16, campo15, font_style2)
                ws.write(row_num, 17, campo16, font_style2)
                ws.write(row_num, 18, campo18, font_style2)
                ws.write(row_num, 19, campo19, style1)
                ws.write(row_num, 20, campo20, style1)
                ws.write(row_num, 21, campo22, font_style2)
                row_num += 1
            wb.save(directory)
            usernotify = User.objects.get(pk=request.user.pk)
            pers = Persona.objects.get(usuario=usernotify)
            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.cuerpo = 'Reporte Listo'
                noti.url = "{}reportes/horarios/{}".format(MEDIA_URL, nombre_archivo)
                noti.save()
            else:
                noti = Notificacion(cuerpo='Reporte Listo', titulo='Reporte de horarios, aulas y actividades',
                                    destinatario=pers, url="{}reportes/horarios/{}".format(MEDIA_URL, nombre_archivo),
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)
                noti.save(request)
            try:
                send_user_notification(user=usernotify, payload={
                    "head": noti.titulo,
                    "body": noti.cuerpo,
                    "action": "notificacion",
                    "timestamp": time.mktime(datetime.now().timetuple()),
                    "url": noti.url,
                    "btn_notificaciones": traerNotificaciones(request, data, pers),
                    "mensaje": 'Su reporte ha sido generado con exito'
                }, ttl=500)
            except:
                pass
        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)
            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.cuerpo = 'Error en el reporte {} - linea {}'.format(ex, sys.exc_info()[-1].tb_lineno)
                noti.url = ""
                noti.save()
            else:
                noti = Notificacion(cuerpo='Error en el reporte {} - linea {}'.format(ex, sys.exc_info()[-1].tb_lineno),
                                    titulo='Reporte de horarios, aulas y actividades',
                                    destinatario=pers, url="",
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)
                noti.save(request)
            try:
                send_user_notification(user=usernotify, payload={
                    "head": noti.titulo,
                    "body": noti.cuerpo,
                    "action": "notificacion",
                    "timestamp": time.mktime(datetime.now().timetuple()),
                    "url": noti.url,
                    "btn_notificaciones": traerNotificaciones(request, data, pers),
                    "mensaje": textoerror
                }, ttl=500)
            except:
                pass


class reporte_asignaturas_alumnos(threading.Thread):
    def __init__(self, request, data, notiid, periodo):
        self.request = request
        self.data = data
        self.notiid = notiid
        self.periodo = periodo
        threading.Thread.__init__(self)

    def run(self):
        from sga.models import Persona, Notificacion
        try:
            directory = os.path.join(MEDIA_ROOT, 'reportes', 'alumnos')
            request, data, notiid = self.request, self.data, self.notiid
            os.makedirs(directory, exist_ok=True)
            nombre_archivo = "reporte_alumnos_asignaturas_matriculados_{}.xlsx".format(random.randint(1, 10000).__str__())
            directory = os.path.join(MEDIA_ROOT, 'reportes', 'alumnos', nombre_archivo)
            __author__ = 'Unemi'
            wb = openxl.Workbook()
            ws = wb.active
            style_title = openxlFont(name='Arial', size=16, bold=True)
            style_cab = openxlFont(name='Arial', size=10, bold=True)
            alinear = alin(horizontal="center", vertical="center")
            ws.merge_cells('A1:H1')
            ws.merge_cells('A2:H2')
            ws['A1'] = 'UNIVERSIDAD ESTATAL DE MILAGRO'
            ws['A2'] = 'REPORTE DE ALUMNOS MATRICULADOS'
            celda1 = ws['A2']
            celda1.font = style_title
            celda1.alignment = alinear
            columns = ["FACULTAD", "CARRERA", "SECCION", "NIVEL", 'ASIGNATURA', "CEDULA ALUMN.", "ALUMNO",
                       "SEXO", "LGTBI", "GRUPO SOCIOECONOMICO", "ESTADO CIVIL", 'ETNIA', "DISCAPACIDAD", "PORCENTAJE",
                       "CARNET", "VERIFICADO", "BECA", "NOTA FINAL", 'ASISTENCIA', "ESTADO", "CEDULA DOC.",
                       "DOCENTE", "SEXO", "TIPO DOCENTE", "DEDICACION DOCENTE", 'TELEFONO', "EMAIL", "EMAIL INST",
                       "CIUDAD", "DIRECCION", "PARALELO", "VECES DE MATRICULA", 'CREDITOS', "TIPO MATRICULA",
                       "GRATUIDAD",
                       "CARRERA DE LA ASIGNATURA", "CODIGO MATERIA"
                       ]

            row_num = 3
            for col_num in range(0, len(columns)):
                celda = ws.cell(row=row_num, column=(col_num + 1), value=columns[col_num])
                celda.font = style_cab
            row_num = 4

            cursor = connections['sga_select'].cursor()
            sql = "   SET statement_timeout='20000 s'; " \
                  "   SELECT DISTINCT " \
                  "   sga_coordinacion.nombre as sga_coordinacion_nombre, " \
                  "   sga_carrera.nombre AS sga_carrera_nombre, " \
                  "   sga_sesion.nombre AS sga_sesion_nombre, " \
                  "   sga_nivelmalla.nombre as sga_nivelmalla_nombre, " \
                  "   sga_asignatura.nombre as sga_asignatura_nombre, " \
                  "   sga_persona.apellido1 || ' ' || sga_persona.apellido2  || ' ' ||  sga_persona.nombres AS sga_persona_nombres, " \
                  "   sga_materiaasignada.notafinal as notafinal, " \
                  "   sga_materiaasignada.asistenciafinal as asistenciafinal, " \
                  "   (select p.apellido1 || ' ' || p.apellido2 || ' ' || p.nombres from sga_persona p where p.id=sga_profesor.persona_id) as docente, " \
                  "   sga_tipoestado.nombre as estado, " \
                  "   sga_nivelmalla.id as sga_nivelmalla_id," \
                  "   sga_persona.telefono, sga_persona.email, sga_persona.emailinst, sga_persona.ciudad, " \
                  "   sga_persona.direccion || ', ' || sga_persona.direccion2 || ' #:' || sga_persona.num_direccion || ' , SECTOR:' || sga_persona.sector as direccion, sga_materia.paralelo, sga_materiaasignada.matriculas, sga_asignaturamalla.creditos," \
                  "   (select p.cedula from sga_persona p where p.id=sga_profesor.persona_id) as cedula, " \
                  "   (select s.nombre from sga_persona p INNER JOIN sga_sexo s on s.id=p.sexo_id where p.id=sga_profesor.persona_id) as sexo, " \
                  "   (select tp.nombre from sga_tipoprofesor tp where tp.id=sga_profesormateria.tipoprofesor_id) as tipoprofesor, " \
                  "   (select de.nombre from sga_tiempodedicaciondocente de where de.id=sga_profesor.dedicacion_id) as dedicacion, " \
                  "   sga_persona.cedula, (select s.nombre from sga_sexo s where s.id=sga_persona.sexo_id ) as sexoalu , sga_persona.lgtbi, " \
                  "   (select gr.nombre from socioecon_fichasocioeconomicainec f inner join socioecon_gruposocioeconomico gr on gr.id=f.grupoeconomico_id where persona_id=sga_persona.id) as socioeco, " \
                  "   (select r.nombre from sga_perfilinscripcion pa inner join sga_raza r on r.id=pa.raza_id where persona_id=sga_persona.id) as etnia, " \
                  "   (select d.nombre from sga_perfilinscripcion pa inner join sga_discapacidad d on d.id=pa.tipodiscapacidad_id where persona_id=sga_persona.id) as discapacidad, " \
                  "   (select pe.porcientodiscapacidad from sga_perfilinscripcion pe where persona_id=sga_persona.id) as porcentaje, " \
                  "   (select pe.carnetdiscapacidad from sga_perfilinscripcion pe where persona_id=sga_persona.id) as carnetdiscapacidad, " \
                  "   (select pe.verificadiscapacidad from sga_perfilinscripcion pe where persona_id=sga_persona.id) as verificadiscapacidad, " \
                  "   (select be.id from sga_inscripcionbecario be where be.inscripcion_id=sga_inscripcion.id) as tienebeca, " \
                  "   (select pe.nombre from med_personaextension es inner join sga_personaestadocivil pe on pe.id=es.estadocivil_id where es.persona_id=sga_persona.id) as estadocivil, " \
                  " (case gru.tipomatricula when 1 then 'REGULAR'  when 2 then 'IREGULAR' else '' end) as tipomatricula, " \
                  " (case gru.estado_gratuidad when 1 then 'GRATUIDAD COMPLETA'  when 2 then 'GRATUIDAD PARCIAL' when 3 then 'PERDIDA DE GRATUIDAD' else '' end) as estado_gratuidad," \
                  "   (select carr.nombre from sga_asignaturamalla asig inner join sga_malla mall on mall.id=asig.malla_id inner join sga_carrera carr on carr.id=mall.carrera_id where asig.id=sga_asignaturamalla.id) as Carrera_de_la_asignatura," \
                  "sga_materia.id as id " \
                  " FROM sga_persona sga_persona " \
                  "      RIGHT OUTER JOIN sga_inscripcion sga_inscripcion ON sga_persona.id = sga_inscripcion.persona_id " \
                  "     INNER JOIN sga_matricula sga_matricula ON sga_matricula.inscripcion_id=sga_inscripcion.id " \
                  " left JOIN sga_matriculagruposocioeconomico gru ON gru.matricula_id=sga_matricula.id " \
                  "    inner join sga_materiaasignada sga_materiaasignada on sga_materiaasignada.matricula_id=sga_matricula.id " \
                  "     inner join sga_materia sga_materia on sga_materia.id=sga_materiaasignada.materia_id " \
                  "     LEFT join sga_profesormateria on sga_profesormateria.materia_id=sga_materia.id " \
                  "     LEFT join sga_profesor on sga_profesor.id=sga_profesormateria.profesor_id " \
                  "     inner join sga_asignatura sga_asignatura on sga_asignatura.id=sga_materia.asignatura_id " \
                  "      inner join sga_asignaturamalla sga_asignaturamalla on sga_asignaturamalla.id=sga_materia.asignaturamalla_id " \
                  "    inner join sga_nivel sga_nivel ON sga_nivel.id=sga_matricula.nivel_id and sga_nivel.periodo_id= '" + str(self.periodo) + "' " \
                                   "      inner join sga_nivelmalla sga_nivelmalla on sga_nivelmalla.id=sga_asignaturamalla.nivelmalla_id " \
                                   "     INNER JOIN sga_carrera sga_carrera ON sga_inscripcion.carrera_id = sga_carrera.id " \
                                   "     INNER JOIN sga_coordinacion_carrera on sga_coordinacion_carrera.carrera_id=sga_carrera.id " \
                                   "     INNER JOIN sga_coordinacion on sga_coordinacion.id=sga_coordinacion_carrera.coordinacion_id " \
                                   "     INNER JOIN sga_modalidad sga_modalidad ON sga_inscripcion.modalidad_id = sga_modalidad.id " \
                                   "     INNER JOIN sga_sesion sga_sesion ON sga_inscripcion.sesion_id = sga_sesion.id " \
                                   "     inner join sga_tipoestado on sga_tipoestado.id=sga_materiaasignada.estado_id " \
                                   "    where sga_matricula.estado_matricula in (2,3) and sga_matricula.id not in (select ret.matricula_id from sga_retiromatricula ret) and sga_profesormateria.activo=True and sga_profesormateria.id = (select pm.id from sga_profesormateria pm inner join sga_profesor profe on pm.profesor_id=profe.id inner join sga_persona p on p.id=profe.persona_id  where pm.materia_id=sga_materiaasignada.materia_id and pm.status=True and pm.activo=True and (pm.tipoprofesor_id in (1, 3, 8, 14, 16)) order by pm.tipoprofesor_id LIMIT 1) order by sga_carrera.nombre,sga_nivelmalla.id,sga_sesion.nombre,sga_persona_nombres"
            cursor.execute(sql)
            results = cursor.fetchall()
            for r in results:
                campo1 = u"%s" % r[0]
                campo2 = u"%s" % r[1]
                campo3 = u"%s" % r[2]
                campo4 = u"%s" % r[3]
                campo5 = u"%s" % r[4]
                campo6 = u"%s" % r[5]
                campo7 = float(r[6]) if r[6] else 0
                campo8 = int(r[7]) if r[7] else 0
                campo9 = u"%s" % r[8]
                campo10 = u"%s" % r[9]
                campo33 = u"%s" % r[33]
                campo11 = u"%s" % r[11]
                campo12 = u"%s" % r[12]
                campo13 = u"%s" % r[13]
                campo14 = u"%s" % r[14]
                campo15 = u"%s" % r[15]
                campo16 = u"%s" % r[16]
                campo17 = int(r[17]) if r[17] else 0
                campo18 = float(r[18]) if r[18] else 0
                campo19 = u"%s" % r[19]
                campo20 = u"%s" % r[20]
                campo21 = u"%s" % r[21]
                campo22 = u"%s" % r[22]
                campo23 = u"%s" % r[23]
                campo24 = u"%s" % r[24]
                campo25 = "SI" if r[25] else "NO"
                campo26 = u"%s" % r[26]
                campo27 = u"%s" % r[27]
                campo28 = u"%s" % r[28] if r[28] else "NINGUNA"
                campo29 = float(r[29]) if r[29] else 0
                campo30 = u"%s" % r[30]
                campo31 = "SI" if r[31] else "NO"
                campo32 = "SI" if r[31] else "NO"
                campo34 = u"%s" % r[34]
                campo35 = u"%s" % r[35]
                campo36 = u"%s" % r[36]
                campo37 = int(r[37]) if r[37] else 0
                #  ws.cell(row=row_num, column=1, value=numero)
                ws.cell(row=row_num, column=1, value=str(campo1))
                ws.cell(row=row_num, column=2, value=str(campo2))
                ws.cell(row=row_num, column=3, value=str(campo3))
                ws.cell(row=row_num, column=4, value=str(campo4))
                ws.cell(row=row_num, column=5, value=str(campo5))
                ws.cell(row=row_num, column=6, value=str(campo23))
                ws.cell(row=row_num, column=7, value=str(campo6))
                ws.cell(row=row_num, column=8, value=str(campo24))
                ws.cell(row=row_num, column=9, value=str(campo25))
                ws.cell(row=row_num, column=10, value=str(campo26))
                ws.cell(row=row_num, column=11, value=str(campo33))
                ws.cell(row=row_num, column=12, value=str(campo27))
                ws.cell(row=row_num, column=13, value=str(campo28))
                ws.cell(row=row_num, column=14, value=campo29)
                ws.cell(row=row_num, column=15, value=str(campo30))
                ws.cell(row=row_num, column=16, value=str(campo31))
                ws.cell(row=row_num, column=17, value=str(campo32))
                ws.cell(row=row_num, column=18, value=campo7)
                ws.cell(row=row_num, column=19, value=campo8)
                ws.cell(row=row_num, column=20, value=str(campo10))
                ws.cell(row=row_num, column=21, value=str(campo19))
                ws.cell(row=row_num, column=22, value=str(campo9))
                ws.cell(row=row_num, column=23, value=str(campo20))
                ws.cell(row=row_num, column=24, value=str(campo21))
                ws.cell(row=row_num, column=25, value=str(campo22))
                ws.cell(row=row_num, column=26, value=str(campo11))
                ws.cell(row=row_num, column=27, value=str(campo12))
                ws.cell(row=row_num, column=28, value=str(campo13))
                ws.cell(row=row_num, column=29, value=str(campo14))
                ws.cell(row=row_num, column=30, value=str(campo15))
                ws.cell(row=row_num, column=31, value=str(campo16))
                ws.cell(row=row_num, column=32, value=campo17)
                ws.cell(row=row_num, column=33, value=campo18)
                ws.cell(row=row_num, column=34, value=str(campo34))
                ws.cell(row=row_num, column=35, value=str(campo35))
                ws.cell(row=row_num, column=36, value=str(campo36))
                ws.cell(row=row_num, column=37, value=campo37)

                row_num += 1
            wb.save(directory)

            user = User.objects.get(pk=request.user.pk)
            pers = Persona.objects.get(usuario=user)
            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.cuerpo = 'Excel listo'
                noti.url = "{}reportes/alumnos/{}".format(MEDIA_URL, nombre_archivo)
                noti.save()
            else:
                noti = Notificacion(cuerpo='Reporte listo', titulo='Reporte de asignaturas y estudiantes',
                                    destinatario=pers, url="{}reportes/alumnos/{}".format(MEDIA_URL, nombre_archivo),
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)
                noti.save(request)

            try:
                send_user_notification(user=user, payload={
                    "head": noti.titulo,
                    "body": noti.cuerpo,
                    "action": "notificacion",
                    "timestamp": time.mktime(datetime.now().timetuple()),
                    "url": noti.url,
                    "btn_notificaciones": traerNotificaciones(request, data, pers),
                    "mensaje": 'Su reporte ha sido generado con exito'
                }, ttl=500)
            except:
                pass

        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)
            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.cuerpo = 'Error en el reporte {} - linea {}'.format(ex, sys.exc_info()[-1].tb_lineno)
                noti.url = ""
                noti.save()
            else:
                noti = Notificacion(cuerpo='Error en el reporte {} - linea {}'.format(ex, sys.exc_info()[-1].tb_lineno),
                                    titulo='Reporte de asignaturas y estudiantes',
                                    destinatario=pers, url="",
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)
                noti.save(request)
            try:
                send_user_notification(user=user, payload={
                    "head": noti.titulo,
                    "body": noti.cuerpo,
                    "action": "notificacion",
                    "timestamp": time.mktime(datetime.now().timetuple()),
                    "url": noti.url,
                    "btn_notificaciones": traerNotificaciones(request, data, pers),
                    "mensaje": textoerror
                }, ttl=500)
            except:
                pass

class reporte_asignaturas_alumnos_matriculados(threading.Thread):
    def __init__(self, request, data, notiid, periodo):
        self.request = request
        self.data = data
        self.notiid = notiid
        self.periodo = periodo
        threading.Thread.__init__(self)

    def run(self):
        from sga.models import Persona, Notificacion
        try:
            directory = os.path.join(MEDIA_ROOT, 'reportes', 'alumnos')
            request, data, notiid, periodo = self.request, self.data, self.notiid, self.periodo
            user = User.objects.get(pk=request.user.pk)
            ePersona = Persona.objects.get(usuario=user)
            os.makedirs(directory, exist_ok=True)
            nombre_archivo = "reporte_alumnos_asignaturas_matriculados_{}.xlsx".format(random.randint(1, 10000).__str__())
            directory = os.path.join(MEDIA_ROOT, 'reportes', 'alumnos', nombre_archivo)
            workbook = xlsxwriter.Workbook(directory, {'constant_memory': True})
            ws = workbook.add_worksheet('listado')
            titulo = workbook.add_format(
                {'align': 'center', 'valign': 'vcenter', 'font_name': 'Times New Roman', 'bold': True,
                 'font_color': 'blue', 'font_size': 18})
            font_style = workbook.add_format({'bold': True, 'border': 1, 'align': 'center', 'valign': 'vcenter'})
            font_style2 = workbook.add_format({'font_name': 'Arial', 'font_size': 10, 'bold': False})
            ws.merge_range(0, 0, 0, 7, 'UNIVERSIDAD ESTATAL DE MILAGRO', titulo)
            columns = [
                (u"FACULTAD", 40),
                (u"CARRERA", 40),
                (u"SECCION", 15),
                (u"NIVEL", 15),
                (u"ASIGNATURA", 30),
                (u"CEDULA", 15),
                (u"ALUMNO", 40),
                (u"NOTA FINAL", 10),
                (u"PROMEDIO", 10),
                (u"ASISTENCIA", 10),
                (u"ESTADO", 10),
                (u"DOCENTE", 40),
                (u"TELEFONO", 10),
                (u"EMAIL", 25),
                (u"EMAIL INST", 25),
                (u"CIUDAD", 20),
                (u"DIRECCION", 50),
                (u"PARALELO", 15)
            ]
            row_num = 3
            for col_num in range(len(columns)):
                ws.write(row_num, col_num, columns[col_num][0], font_style)
                ws.set_column(col_num, col_num, columns[col_num][1])
            cursor = connections['sga_select'].cursor()
            # QUERY SQL -- No tocar
            sql = f"""
                        SET statement_timeout='20000 s';
                        SELECT DISTINCT
                        sga_coordinacion.nombre as sga_coordinacion_nombre,
                        sga_carrera.nombre AS sga_carrera_nombre,
                        sga_sesion.nombre AS sga_sesion_nombre,
                        sga_nivelmalla.nombre as sga_nivelmalla_nombre,
                        sga_asignatura.nombre as sga_asignatura_nombre,
                        sga_persona.apellido1 || ' ' || sga_persona.apellido2  || ' ' ||  sga_persona.nombres AS sga_persona_nombres,
                        sga_materiaasignada.notafinal as notafinal,
                        sga_materiaasignada.asistenciafinal as asistenciafinal,
                        (select p.apellido1 || ' ' || p.apellido2 || ' ' || p.nombres from sga_persona p where p.id=sga_profesor.persona_id) as docente,
                        sga_tipoestado.nombre as estado,
                        sga_nivelmalla.id as sga_nivelmalla_id,
                        sga_persona.telefono, sga_persona.email, sga_persona.emailinst, sga_persona.ciudad,
                        sga_persona.direccion || ', ' || sga_persona.direccion2 || ' #:' || sga_persona.num_direccion || ' , SECTOR:' || sga_persona.sector as direccion, 
                        sga_materia.paralelo, 
                        (select sum((case when a1.id = '1939' then(ma1.notafinal * 0.15) else (case when mo1.id = 4 then(ma1.notafinal * 0.10) else (ma1.notafinal * 0.25) end) end) )
                        from sga_materiaasignada ma1, sga_materia m1, sga_modeloevaluativo mo1, sga_asignatura a1 where ma1.matricula_id = sga_matricula.id and ma1.status = true and m1.id = ma1.materia_id and m1.status = true
                        and mo1.id = m1.modeloevaluativo_id and mo1.status = true and a1.id = m1.asignatura_id and a1.status = true) as notapromedio, sga_persona.cedula
                        FROM sga_persona sga_persona
                        RIGHT OUTER JOIN sga_inscripcion sga_inscripcion ON sga_persona.id = sga_inscripcion.persona_id
                        INNER JOIN sga_matricula sga_matricula ON sga_matricula.inscripcion_id=sga_inscripcion.id
                        inner join sga_materiaasignada sga_materiaasignada on sga_materiaasignada.matricula_id=sga_matricula.id
                        inner join sga_materia sga_materia on sga_materia.id=sga_materiaasignada.materia_id
                        LEFT join sga_profesormateria on sga_profesormateria.materia_id=sga_materia.id
                        LEFT join sga_profesor on sga_profesor.id=sga_profesormateria.profesor_id
                        inner join sga_asignatura sga_asignatura on sga_asignatura.id=sga_materia.asignatura_id
                        inner join sga_asignaturamalla sga_asignaturamalla on sga_asignaturamalla.id=sga_materia.asignaturamalla_id
                        inner join sga_nivel sga_nivel ON sga_nivel.id=sga_matricula.nivel_id and sga_nivel.periodo_id= {periodo}
                        inner join sga_nivelmalla sga_nivelmalla on sga_nivelmalla.id=sga_asignaturamalla.nivelmalla_id
                        INNER JOIN sga_carrera sga_carrera ON sga_inscripcion.carrera_id = sga_carrera.id
                        INNER JOIN sga_coordinacion_carrera on sga_coordinacion_carrera.carrera_id=sga_carrera.id 
                        INNER JOIN sga_coordinacion on sga_coordinacion.id=sga_coordinacion_carrera.coordinacion_id
                        INNER JOIN sga_modalidad sga_modalidad ON sga_inscripcion.modalidad_id = sga_modalidad.id
                        INNER JOIN sga_sesion sga_sesion ON sga_inscripcion.sesion_id = sga_sesion.id
                        inner join sga_tipoestado on sga_tipoestado.id=sga_materiaasignada.estado_id
                        where sga_matricula.id not in (select ma.id from (select mat.id as id, count(ma.materia_id)  as numero 
                        from sga_Matricula mat, sga_Nivel n, sga_materiaasignada  ma, sga_materia mate, sga_asignatura asi
                        where mat.nivel_id = n.id and mat.id = ma.matricula_id and ma.materia_id = mate.id and mate.asignatura_id = asi.id and n.periodo_id = {periodo}
                        and asi.modulo = True group by mat.id)  ma, (select mat.id as id, count(ma.materia_id) as numero from sga_Matricula mat, sga_Nivel n, sga_materiaasignada ma,
                        sga_materia mate, sga_asignatura asi where mat.nivel_id = n.id and mat.id = ma.matricula_id and ma.materia_id = mate.id and mate.asignatura_id = asi.id and n.periodo_id = {periodo}
                        group by mat.id) mo where ma.id = mo.id and ma.numero = mo.numero) order by sga_carrera.nombre,sga_nivelmalla.id,sga_sesion.nombre,sga_persona_nombres
                    """
            cursor.execute(sql)
            results = cursor.fetchall()
            row_num = 4
            for r in results:
                ws.write(row_num, 0, r[0], font_style2)
                ws.write(row_num, 1, r[1], font_style2)
                ws.write(row_num, 2, r[2], font_style2)
                ws.write(row_num, 3, r[3], font_style2)
                ws.write(row_num, 4, r[4], font_style2)
                ws.write(row_num, 5, r[18], font_style2)
                ws.write(row_num, 6, r[5], font_style2)
                ws.write(row_num, 7, r[6], font_style2)
                ws.write(row_num, 8, round(float(r[17])) if r[17] else float(0), font_style2)
                ws.write(row_num, 9, r[7], font_style2)
                ws.write(row_num, 10, r[9], font_style2)
                ws.write(row_num, 11, r[8], font_style2)
                ws.write(row_num, 12, r[11], font_style2)
                ws.write(row_num, 13, r[12], font_style2)
                ws.write(row_num, 14, r[13], font_style2)
                ws.write(row_num, 15, r[14], font_style2)
                ws.write(row_num, 16, r[15], font_style2)
                ws.write(row_num, 17, r[16], font_style2)
                row_num += 1

            workbook.close()

            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.cuerpo = 'Excel listo'
                noti.url = "{}reportes/alumnos/{}".format(MEDIA_URL, nombre_archivo)
                noti.save()
            else:
                noti = Notificacion(cuerpo='Reporte listo',
                                    titulo='Reporte de aptos EHEP',
                                    destinatario=ePersona,
                                    url="{}reportes/alumnos/{}".format(MEDIA_URL, nombre_archivo),
                                    prioridad=1,
                                    app_label='SGA',
                                    fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2,
                                    en_proceso=False)
                noti.save(request)

            try:
                send_user_notification(user=user, payload={
                    "head": noti.titulo,
                    "body": noti.cuerpo,
                    "action": "notificacion",
                    "timestamp": time.mktime(datetime.now().timetuple()),
                    "url": noti.url,
                    "btn_notificaciones": traerNotificaciones(request, data, ePersona),
                    "mensaje": 'Su reporte ha sido generado con exito'
                }, ttl=500)
            except:
                pass
        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)
            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.cuerpo = 'Error en el reporte {} - linea {}'.format(ex, sys.exc_info()[-1].tb_lineno)
                noti.url = ""
                noti.save()
            else:
                noti = Notificacion(cuerpo='Error en el reporte {} - linea {}'.format(ex, sys.exc_info()[-1].tb_lineno),
                                    titulo=f'Reporte de aptos EHEP',
                                    destinatario=ePersona, url="",
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)
                noti.save(request)
            try:
                send_user_notification(user=user, payload={
                    "head": noti.titulo,
                    "body": noti.cuerpo,
                    "action": "notificacion",
                    "timestamp": time.mktime(datetime.now().timetuple()),
                    "url": noti.url,
                    "btn_notificaciones": traerNotificaciones(request, data, ePersona),
                    "mensaje": textoerror
                }, ttl=500)
            except:
                pass


class reporte_asignaturas_alumnos_coordinaciones(threading.Thread):
    def __init__(self, request, data, notiid, periodo):
        self.request = request
        self.data = data
        self.notiid = notiid
        self.periodo = periodo
        threading.Thread.__init__(self)

    def run(self):
        from sga.models import Persona, Notificacion
        try:
            directory = os.path.join(MEDIA_ROOT, 'reportes', 'alumnos')
            request, data, notiid, periodo = self.request, self.data, self.notiid, self.periodo
            user = User.objects.get(pk=request.user.pk)
            ePersona = Persona.objects.get(usuario=user)
            os.makedirs(directory, exist_ok=True)
            nombre_archivo = "reporte_alumnos_asignaturas_matriculados_coordinaciones_{}.xlsx".format(random.randint(1, 10000).__str__())
            directory = os.path.join(MEDIA_ROOT, 'reportes', 'alumnos', nombre_archivo)
            workbook = xlsxwriter.Workbook(directory, {'constant_memory': True})
            ws = workbook.add_worksheet('Hoja1')
            ws.set_column(0, 37, 30)
            formatoceldaleft = workbook.add_format({'text_wrap': True, 'align': 'left'})
            decimalformat = workbook.add_format({'num_format': '#,##0.00'})
            ws.write('A1', 'FACULTAD')
            ws.write('B1', 'CARRERA')
            ws.write('C1', 'SECCION')
            ws.write('D1', 'NIVEL')
            ws.write('E1', 'ASIGNATURA')
            ws.write('F1', 'CEDULA ALUMN.')
            ws.write('G1', 'ALUMNO')
            ws.write('H1', 'SEXO')
            ws.write('I1', 'LGTBI')
            ws.write('J1', 'GRUPO SOCIOECONOMICO')
            ws.write('K1', 'ESTADO CIVIL')
            ws.write('L1', 'ETNIA')
            ws.write('M1', 'DISCAPACIDAD')
            ws.write('N1', 'PORCENTAJE')
            ws.write('O1', 'CARNET')
            ws.write('P1', 'VERIFICADO')
            ws.write('Q1', 'BECA')
            ws.write('R1', 'NOTA FINAL')
            ws.write('S1', 'ASISTENCIA')
            ws.write('T1', 'ESTADO')
            ws.write('U1', 'CEDULA DOC.')
            ws.write('V1', 'DOCENTE')
            ws.write('W1', 'SEXO')
            ws.write('X1', 'TIPO DOCENTE')
            ws.write('Y1', 'DEDICACION DOCENTE')
            ws.write('Z1', 'TELEFONO')
            ws.write('AA1', 'EMAIL')
            ws.write('AB1', 'EMAIL INST')
            ws.write('AC1', 'CIUDAD')
            ws.write('AD1', 'DIRECCION')
            ws.write('AE1', 'PARALELO')
            ws.write('AF1', 'VECES DE MATRICULA')
            ws.write('AG1', 'CREDITOS')
            ws.write('AH1', 'TIPO MATRICULA')
            ws.write('AI1', 'GRATUIDAD')
            ws.write('AJ1', 'CARRERA DE LA ASIGNATURA')
            ws.write('AK1', 'CODIGO MATERIA')

            coordinaciones_ = tuple(ePersona.mis_coordinaciones().values_list('id', flat=True))
            cursor = connections['sga_select'].cursor()
            sql = "   SET statement_timeout='20000 s'; " \
                  "   SELECT DISTINCT " \
                  "   sga_coordinacion.nombre as sga_coordinacion_nombre, " \
                  "   sga_carrera.nombre AS sga_carrera_nombre, " \
                  "   sga_sesion.nombre AS sga_sesion_nombre, " \
                  "   sga_nivelmalla.nombre as sga_nivelmalla_nombre, " \
                  "   sga_asignatura.nombre as sga_asignatura_nombre, " \
                  "   sga_persona.apellido1 || ' ' || sga_persona.apellido2  || ' ' ||  sga_persona.nombres AS sga_persona_nombres, " \
                  "   sga_materiaasignada.notafinal as notafinal, " \
                  "   sga_materiaasignada.asistenciafinal as asistenciafinal, " \
                  "   (select p.apellido1 || ' ' || p.apellido2 || ' ' || p.nombres from sga_persona p where p.id=sga_profesor.persona_id) as docente, " \
                  "   sga_tipoestado.nombre as estado, " \
                  "   sga_nivelmalla.id as sga_nivelmalla_id," \
                  "   sga_persona.telefono, sga_persona.email, sga_persona.emailinst, sga_persona.ciudad, " \
                  "   sga_persona.direccion || ', ' || sga_persona.direccion2 || ' #:' || sga_persona.num_direccion || ' , SECTOR:' || sga_persona.sector as direccion, sga_materia.paralelo, sga_materiaasignada.matriculas, sga_asignaturamalla.creditos," \
                  "   (select p.cedula from sga_persona p where p.id=sga_profesor.persona_id) as cedula, " \
                  "   (select s.nombre from sga_persona p INNER JOIN sga_sexo s on s.id=p.sexo_id where p.id=sga_profesor.persona_id) as sexo, " \
                  "   (select tp.nombre from sga_tipoprofesor tp where tp.id=sga_profesormateria.tipoprofesor_id) as tipoprofesor, " \
                  "   (select de.nombre from sga_tiempodedicaciondocente de where de.id=sga_profesor.dedicacion_id) as dedicacion, " \
                  "   sga_persona.cedula, (select s.nombre from sga_sexo s where s.id=sga_persona.sexo_id ) as sexoalu , sga_persona.lgtbi, " \
                  "   (select gr.nombre from socioecon_fichasocioeconomicainec f inner join socioecon_gruposocioeconomico gr on gr.id=f.grupoeconomico_id where persona_id=sga_persona.id) as socioeco, " \
                  "   (select r.nombre from sga_perfilinscripcion pa inner join sga_raza r on r.id=pa.raza_id where persona_id=sga_persona.id) as etnia, " \
                  "   (select d.nombre from sga_perfilinscripcion pa inner join sga_discapacidad d on d.id=pa.tipodiscapacidad_id where persona_id=sga_persona.id) as discapacidad, " \
                  "   (select pe.porcientodiscapacidad from sga_perfilinscripcion pe where persona_id=sga_persona.id) as porcentaje, " \
                  "   (select pe.carnetdiscapacidad from sga_perfilinscripcion pe where persona_id=sga_persona.id) as carnetdiscapacidad, " \
                  "   (select pe.verificadiscapacidad from sga_perfilinscripcion pe where persona_id=sga_persona.id) as verificadiscapacidad, " \
                  "   (select be.id from sga_inscripcionbecario be where be.inscripcion_id=sga_inscripcion.id) as tienebeca, " \
                  "   (select pe.nombre from med_personaextension es inner join sga_personaestadocivil pe on pe.id=es.estadocivil_id where es.persona_id=sga_persona.id) as estadocivil, " \
                  " (case gru.tipomatricula when 1 then 'REGULAR'  when 2 then 'IREGULAR' else '' end) as tipomatricula, " \
                  " (case gru.estado_gratuidad when 1 then 'GRATUIDAD COMPLETA'  when 2 then 'GRATUIDAD PARCIAL' when 3 then 'PERDIDA DE GRATUIDAD' else '' end) as estado_gratuidad," \
                  "   (select carr.nombre from sga_asignaturamalla asig inner join sga_malla mall on mall.id=asig.malla_id inner join sga_carrera carr on carr.id=mall.carrera_id where asig.id=sga_asignaturamalla.id) as Carrera_de_la_asignatura," \
                  "sga_materia.id as id " \
                  " FROM sga_persona sga_persona " \
                  "      RIGHT OUTER JOIN sga_inscripcion sga_inscripcion ON sga_persona.id = sga_inscripcion.persona_id " \
                  "     INNER JOIN sga_matricula sga_matricula ON sga_matricula.inscripcion_id=sga_inscripcion.id " \
                  " left JOIN sga_matriculagruposocioeconomico gru ON gru.matricula_id=sga_matricula.id " \
                  "    inner join sga_materiaasignada sga_materiaasignada on sga_materiaasignada.matricula_id=sga_matricula.id " \
                  "     inner join sga_materia sga_materia on sga_materia.id=sga_materiaasignada.materia_id " \
                  "     LEFT join sga_profesormateria on sga_profesormateria.materia_id=sga_materia.id " \
                  "     LEFT join sga_profesor on sga_profesor.id=sga_profesormateria.profesor_id " \
                  "     inner join sga_asignatura sga_asignatura on sga_asignatura.id=sga_materia.asignatura_id " \
                  "      inner join sga_asignaturamalla sga_asignaturamalla on sga_asignaturamalla.id=sga_materia.asignaturamalla_id " \
                  "    inner join sga_nivel sga_nivel ON sga_nivel.id=sga_matricula.nivel_id and sga_nivel.periodo_id= '" + str(periodo) + "' " \
                                                                                                                                      "      inner join sga_nivelmalla sga_nivelmalla on sga_nivelmalla.id=sga_asignaturamalla.nivelmalla_id " \
                                                                                                                                      "     INNER JOIN sga_carrera sga_carrera ON sga_inscripcion.carrera_id = sga_carrera.id " \
                                                                                                                                      "     INNER JOIN sga_coordinacion_carrera on sga_coordinacion_carrera.carrera_id=sga_carrera.id " \
                                                                                                                                      "     INNER JOIN sga_coordinacion on sga_coordinacion.id=sga_coordinacion_carrera.coordinacion_id " \
                                                                                                                                      "     INNER JOIN sga_modalidad sga_modalidad ON sga_inscripcion.modalidad_id = sga_modalidad.id " \
                                                                                                                                      "     INNER JOIN sga_sesion sga_sesion ON sga_inscripcion.sesion_id = sga_sesion.id " \
                                                                                                                                      "     inner join sga_tipoestado on sga_tipoestado.id=sga_materiaasignada.estado_id " \
                                                                                                                                      "    where sga_matricula.estado_matricula in (2,3) and sga_matricula.id not in (select ret.matricula_id from sga_retiromatricula ret) and sga_profesormateria.activo=True and sga_profesormateria.id = (select pm.id from sga_profesormateria pm inner join sga_profesor profe on pm.profesor_id=profe.id inner join sga_persona p on p.id=profe.persona_id  where pm.materia_id=sga_materiaasignada.materia_id and pm.status=True and pm.activo=True AND sga_coordinacion.id IN " + str(
                coordinaciones_) + " and (pm.tipoprofesor_id=1 or pm.tipoprofesor_id=3) order by pm.tipoprofesor_id LIMIT 1) order by sga_carrera.nombre,sga_nivelmalla.id,sga_sesion.nombre,sga_persona_nombres "
            cursor.execute(sql)
            results = cursor.fetchall()
            row_num = 1
            for r in results:
                i = 0
                campo1 = u"%s" % r[0]
                campo2 = u"%s" % r[1]
                campo3 = u"%s" % r[2]
                campo4 = u"%s" % r[3]
                campo5 = u"%s" % r[4]
                campo6 = u"%s" % r[5]
                campo7 = float(r[6])
                campo8 = int(r[7])
                campo9 = u"%s" % r[8]
                campo10 = u"%s" % r[9]
                campo33 = u"%s" % r[33]
                campo11 = u"%s" % r[11]
                campo12 = u"%s" % r[12]
                campo13 = u"%s" % r[13]
                campo14 = u"%s" % r[14]
                campo15 = u"%s" % r[15]
                campo16 = u"%s" % r[16]
                campo17 = int(r[17])
                campo18 = float(r[18])
                campo19 = u"%s" % r[19]
                campo20 = u"%s" % r[20]
                campo21 = u"%s" % r[21]
                campo22 = u"%s" % r[22]
                campo23 = u"%s" % r[23]
                campo24 = u"%s" % r[24]
                campo25 = "SI" if r[25] else "NO"
                campo26 = u"%s" % r[26]
                campo27 = u"%s" % r[27]
                campo28 = u"%s" % r[28] if r[28] else "NINGUNA"
                campo29 = float(r[29]) if r[29] else float(0)
                campo30 = u"%s" % r[30]
                campo31 = "SI" if r[31] else "NO"
                campo32 = "SI" if r[31] else "NO"
                campo34 = u"%s" % r[34]
                campo35 = u"%s" % r[35]
                campo36 = u"%s" % r[36]
                campo37 = int(r[37]) if r[37] else 0
                ws.write(row_num, 0, str(campo1), formatoceldaleft)
                ws.write(row_num, 1, str(campo2), formatoceldaleft)
                ws.write(row_num, 2, str(campo3), formatoceldaleft)
                ws.write(row_num, 3, str(campo4), formatoceldaleft)
                ws.write(row_num, 4, str(campo5), formatoceldaleft)
                ws.write(row_num, 5, str(campo23), formatoceldaleft)
                ws.write(row_num, 6, str(campo6), formatoceldaleft)
                ws.write(row_num, 7, str(campo24), formatoceldaleft)
                ws.write(row_num, 8, str(campo25), formatoceldaleft)
                ws.write(row_num, 9, str(campo26), formatoceldaleft)
                ws.write(row_num, 10, str(campo33), formatoceldaleft)
                ws.write(row_num, 11, str(campo27), formatoceldaleft)
                ws.write(row_num, 12, str(campo28), formatoceldaleft)
                ws.write(row_num, 13, campo29, decimalformat)
                ws.write(row_num, 14, str(campo30), formatoceldaleft)
                ws.write(row_num, 15, str(campo31), formatoceldaleft)
                ws.write(row_num, 16, str(campo32), formatoceldaleft)
                ws.write(row_num, 17, campo7, decimalformat)
                ws.write(row_num, 18, campo8)
                ws.write(row_num, 19, str(campo10), formatoceldaleft)
                ws.write(row_num, 20, str(campo19), formatoceldaleft)
                ws.write(row_num, 21, str(campo9), formatoceldaleft)
                ws.write(row_num, 22, str(campo20), formatoceldaleft)
                ws.write(row_num, 23, str(campo21), formatoceldaleft)
                ws.write(row_num, 24, str(campo22), formatoceldaleft)
                ws.write(row_num, 25, str(campo11), formatoceldaleft)
                ws.write(row_num, 26, str(campo12), formatoceldaleft)
                ws.write(row_num, 27, str(campo13), formatoceldaleft)
                ws.write(row_num, 28, str(campo14), formatoceldaleft)
                ws.write(row_num, 29, str(campo15), formatoceldaleft)
                ws.write(row_num, 30, str(campo16), formatoceldaleft)
                ws.write(row_num, 31, campo17)
                ws.write(row_num, 32, campo18, decimalformat)
                ws.write(row_num, 33, str(campo34), formatoceldaleft)
                ws.write(row_num, 34, str(campo35), formatoceldaleft)
                ws.write(row_num, 35, str(campo36), formatoceldaleft)
                ws.write(row_num, 36, campo37)
                row_num += 1
            workbook.close()

            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.cuerpo = 'Excel listo'
                noti.url = "{}reportes/alumnos/{}".format(MEDIA_URL, nombre_archivo)
                noti.save()
            else:
                noti = Notificacion(cuerpo='Reporte listo',
                                    titulo='Reporte de asignaturas y estudiantes Filtro: Mis Coordinaciones',
                                    destinatario=ePersona,
                                    url="{}reportes/alumnos/{}".format(MEDIA_URL, nombre_archivo),
                                    prioridad=1,
                                    app_label='SGA',
                                    fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2,
                                    en_proceso=False)
                noti.save(request)

            try:
                send_user_notification(user=user, payload={
                    "head": noti.titulo,
                    "body": noti.cuerpo,
                    "action": "notificacion",
                    "timestamp": time.mktime(datetime.now().timetuple()),
                    "url": noti.url,
                    "btn_notificaciones": traerNotificaciones(request, data, ePersona),
                    "mensaje": 'Su reporte ha sido generado con exito'
                }, ttl=500)
            except:
                pass

        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)
            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.cuerpo = 'Error en el reporte {} - linea {}'.format(ex, sys.exc_info()[-1].tb_lineno)
                noti.url = ""
                noti.save()
            else:
                noti = Notificacion(cuerpo='Error en el reporte {} - linea {}'.format(ex, sys.exc_info()[-1].tb_lineno),
                                    titulo='Reporte de asignaturas y estudiantes Filtro: Mis Coordinaciones',
                                    destinatario=ePersona, url="",
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)
                noti.save(request)
            try:
                send_user_notification(user=user, payload={
                    "head": noti.titulo,
                    "body": noti.cuerpo,
                    "action": "notificacion",
                    "timestamp": time.mktime(datetime.now().timetuple()),
                    "url": noti.url,
                    "btn_notificaciones": traerNotificaciones(request, data, ePersona),
                    "mensaje": textoerror
                }, ttl=500)
            except:
                pass


class reporte_asignaturas_alumnos_sin_profesor(threading.Thread):
    def __init__(self, request, data, notiid, periodo, type):
        self.request = request
        self.data = data
        self.notiid = notiid
        self.periodo = periodo
        self.type = type
        threading.Thread.__init__(self)

    def run(self):
        from sga.models import Persona, Notificacion
        try:
            directory = os.path.join(MEDIA_ROOT, 'reportes', 'alumnos')
            request, data, notiid, periodo, type = self.request, self.data, self.notiid, self.periodo, self.type
            user = User.objects.get(pk=request.user.pk)
            ePersona = Persona.objects.get(usuario=user)
            os.makedirs(directory, exist_ok=True)

            nombre_archivo = "reporte_asignaturas_estudiantes_sin_profesor_{}_{}.xlsx".format(type, random.randint(1, 10000).__str__())
            directory = os.path.join(MEDIA_ROOT, 'reportes', 'alumnos', nombre_archivo)
            workbook = xlsxwriter.Workbook(directory, {'constant_memory': True})
            ws = workbook.add_worksheet('Hoja1')
            merge_format = workbook.add_format({
                'bold': 1,
                'border': 1,
                'align': 'center',
                'valign': 'vcenter'})

            ws.merge_range(0, 0, 0, 7, 'UNIVERSIDAD ESTATAL DE MILAGRO', workbook.add_format({'align': 'center', 'valign': 'vcenter'}))
            columns = [
                (u"FACULTAD", 40),
                (u"CARRERA", 40),
                (u"CEDULA", 25),
                (u"PRIMER APELLIDO", 40),
                (u"SEGUNDO APELLIDO", 40),
                (u"NOMBRES", 40),
                (u"EMAIL", 40),
                (u"EMAIL INSTITUCIONAL", 40),
                (u"CELULAR", 25),
                (u"TELEFONO", 25),
                (u"MATERIA", 40),
                (u"PARALELO", 25),
                (u"NIVEL", 25),
                (u"NOTA FINAL", 40),
                (u"ASISTENCIA FINAL", 40),
                (u"ESTADO", 25),
                (u"NUMERO DE MATRICULA", 25),
                (u"ANIO MALLA", 25),
                (u"FECHA NACIMIENTO", 25),
                (u"MODALIDAD", 25),
                (u"SEXO", 25),
                (u"CANTON RESIDENCIA", 25),
                (u"PROVINCIA", 25),
                (u"PAIS", 25),
                (u"GRUPO SOCIECONOMICO", 25),
            ]
            row_num = 3
            for col_num in range(len(columns)):
                ws.write(row_num, col_num, columns[col_num][0], merge_format)
                ws.set_column(col_num, col_num, columns[col_num][1])
            cursor = connections['default'].cursor()
            where = ""
            if type == 'grado':
                where = "coordinacion.id not in(9, 7) and coordinacion.id in (1,2,3,4,5) and "
            elif type == 'nivelacion':
                where = "coordinacion.id not in(1,2,3,4,5,7) and coordinacion.id in (9) and "
            elif type == 'posgrado':
                where = "coordinacion.id not in(1,2,3,4,5,9) and coordinacion.id in (7) and "
            sql = """SELECT coordinacion.nombre as coordinacion, CASE WHEN carrera.mencion = '' THEN carrera.nombre ELSE (carrera.nombre || ' CON MENCION EN ' || carrera.mencion) END as carrera, 
                                        persona.cedula, persona.apellido1, persona.apellido2, persona.nombres, persona.email, persona.emailinst, persona.telefono, persona.telefono_conv, 
                                        (asignatura.nombre || CASE WHEN materia.alias = '' THEN '' ELSE (' - ('|| materia.alias || ')') END  || CASE WHEN materia.identificacion = '' THEN '' ELSE (' - ['|| materia.identificacion || ']') END) as materia, 
                                        materia.paralelo as paralelo, nivelmalla.nombre as nivel, materiaasig.notafinal as notafinal, materiaasig.asistenciafinal as asistenciafinal, tipoestado.nombre as estado,  materiaasig.matriculas as veces, extract(year from malla.inicio) as aniomalla,  
                                        persona.nacimiento, modalidad.nombre, genero.nombre, canton.nombre, gruposocio.nombre, provincia.nombre, pais.nombre  
                                        from sga_materiaasignada materiaasig 
                                        inner join sga_materia materia on materia.id = materiaasig.materia_id 
                                        inner join sga_nivel nivel on nivel.id = materia.nivel_id 
                                        inner join sga_modalidad modalidad on nivel.modalidad_id = modalidad.id 
                                        inner join sga_asignaturamalla asigmalla on asigmalla.id = materia.asignaturamalla_id 
                                        inner join sga_malla malla on malla.id = asigmalla.malla_id 
                                        inner join sga_carrera carrera on carrera.id = malla.carrera_id 
                                        inner join sga_matricula matricula on matricula.id = materiaasig.matricula_id 
                                        inner join sga_inscripcion inscripcion on inscripcion.id = matricula.inscripcion_id 
                                        inner join sga_persona persona on persona.id = inscripcion.persona_id 
                                        left join sga_sexo genero on genero.id = persona.sexo_id 
                                        left join sga_canton canton on canton.id = persona.canton_id 
                                        left join sga_provincia provincia on provincia.id = persona.provincia_id 
                                        left join sga_pais pais on pais.id = persona.pais_id 
                                        inner join sga_asignatura asignatura on asignatura.id = materia.asignatura_id 
                                        inner join sga_nivelmalla nivelmalla on nivelmalla.id = asigmalla.nivelmalla_id 
                                        inner join sga_tipoestado tipoestado on tipoestado.id = materiaasig.estado_id 
                                        inner join sga_coordinacion_carrera coordinacioncarrera on coordinacioncarrera.carrera_id = carrera.id 
                                        inner join sga_coordinacion coordinacion on coordinacion.id = coordinacioncarrera.coordinacion_id 
                                        left join sga_matriculagruposocioeconomico matri_gruposocio on matri_gruposocio.matricula_id = matricula.id
                                        left join socioecon_gruposocioeconomico gruposocio on matri_gruposocio.gruposocioeconomico_id = gruposocio.id
                                        where %s materia.status=True and materiaasig.status=True AND materiaasig.retiramateria = False
                                          and matricula.status and  nivel.periodo_id = %s""" % (where, periodo)
            cursor.execute(sql)
            results = cursor.fetchall()
            row_num = 4
            for r in results:
                ws.write(row_num, 0, u'%s' % r[0])
                ws.write(row_num, 1, u'%s' % r[1])
                ws.write(row_num, 2, u'%s' % r[2])
                ws.write(row_num, 3, u'%s' % r[3])
                ws.write(row_num, 4, u'%s' % r[4])
                ws.write(row_num, 5, u'%s' % r[5])
                ws.write(row_num, 6, u'%s' % r[6])
                ws.write(row_num, 7, u'%s' % r[7])
                ws.write(row_num, 8, u'%s' % r[8])
                ws.write(row_num, 9, u'%s' % r[9])
                ws.write(row_num, 10, u'%s' % r[10])
                ws.write(row_num, 11, u'%s' % r[11])
                ws.write(row_num, 12, u'%s' % r[12])
                ws.write(row_num, 13, r[13])
                ws.write(row_num, 14, r[14])
                ws.write(row_num, 15, u'%s' % r[15])
                ws.write(row_num, 16, 3 if r[16] > 3 else r[16])
                ws.write(row_num, 17, r[17])
                ws.write(row_num, 18, u'%s' % r[18])
                ws.write(row_num, 19, u'%s' % r[19])
                ws.write(row_num, 20, u'%s' % r[20])
                ws.write(row_num, 21, u'%s' % r[21])
                ws.write(row_num, 22, u'%s' % r[23])
                ws.write(row_num, 23, u'%s' % r[24])
                ws.write(row_num, 24, u'%s' % r[22])
                row_num += 1
            workbook.close()

            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.cuerpo = 'Excel listo'
                noti.url = "{}reportes/alumnos/{}".format(MEDIA_URL, nombre_archivo)
                noti.save()
            else:
                noti = Notificacion(cuerpo='Reporte listo',
                                    titulo=f'Reporte de Asignaturas y estudiantes sin profesor {type}',
                                    destinatario=ePersona,
                                    url="{}reportes/alumnos/{}".format(MEDIA_URL, nombre_archivo),
                                    prioridad=1,
                                    app_label='SGA',
                                    fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2,
                                    en_proceso=False)
                noti.save(request)

            try:
                send_user_notification(user=user, payload={
                    "head": noti.titulo,
                    "body": noti.cuerpo,
                    "action": "notificacion",
                    "timestamp": time.mktime(datetime.now().timetuple()),
                    "url": noti.url,
                    "btn_notificaciones": traerNotificaciones(request, data, ePersona),
                    "mensaje": 'Su reporte ha sido generado con exito'
                }, ttl=500)
            except:
                pass

        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)
            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.cuerpo = 'Error en el reporte {} - linea {}'.format(ex, sys.exc_info()[-1].tb_lineno)
                noti.url = ""
                noti.save()
            else:
                noti = Notificacion(cuerpo='Error en el reporte {} - linea {}'.format(ex, sys.exc_info()[-1].tb_lineno),
                                    titulo=f'Reporte de Asignaturas y estudiantes sin profesor {type}',
                                    destinatario=ePersona, url="",
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)
                noti.save(request)
            try:
                send_user_notification(user=user, payload={
                    "head": noti.titulo,
                    "body": noti.cuerpo,
                    "action": "notificacion",
                    "timestamp": time.mktime(datetime.now().timetuple()),
                    "url": noti.url,
                    "btn_notificaciones": traerNotificaciones(request, data, ePersona),
                    "mensaje": textoerror
                }, ttl=500)
            except:
                pass


class reporte_alumnos_enfermeria_octavo(threading.Thread):
    def __init__(self, request, data, notiid, periodo):
        self.request = request
        self.data = data
        self.notiid = notiid
        self.periodo = periodo
        threading.Thread.__init__(self)

    def run(self):
        from sga.models import Persona, Notificacion, Inscripcion
        try:
            directory = os.path.join(MEDIA_ROOT, 'reportes', 'alumnos')
            request, data, notiid, periodo = self.request, self.data, self.notiid, self.periodo
            user = User.objects.get(pk=request.user.pk)
            ePersona = Persona.objects.get(usuario=user)
            os.makedirs(directory, exist_ok=True)
            nombre_archivo = "reporte_aptos_EHEP_{}.xlsx".format(random.randint(1, 10000).__str__())
            directory = os.path.join(MEDIA_ROOT, 'reportes', 'alumnos', nombre_archivo)
            workbook = xlsxwriter.Workbook(directory, {'constant_memory': True})
            ws = workbook.add_worksheet('Hoja1')
            ws.set_column(0, 37, 30)
            formatoceldaleft = workbook.add_format({'text_wrap': True, 'align': 'left'})
            decimalformat = workbook.add_format({'num_format': '#,##0.00'})
            ws.write('A1', 'CODIGO_IES')
            ws.write('B1', 'NOMBRE_IES')
            ws.write('C1', 'CODIGO_CARRERA')
            ws.write('D1', 'TIPO_IDENTIFICACION')
            ws.write('E1', 'IDENTIFICACION')
            ws.write('F1', 'SEXO')
            ws.write('G1', 'PRIMER_APELLIDO')
            ws.write('H1', 'SEGUNDO_APELLIDO')
            ws.write('I1', 'NOMBRES')
            ws.write('J1', 'FECHA_INICIO_PRIMER_NIVEL')
            ws.write('K1', 'FECHA_INGRESO_CONVALIDACION')
            ws.write('L1', 'FECHA_GRADUACION')
            ws.write('M1', 'FECHA_INICIO')
            ws.write('N1', 'FECHA_FIN')
            ws.write('O1', 'TOTAL_CREDITOS_SIN_TITULACION')
            ws.write('P1', 'NUMERO_CREDITOS_APROBADOS')

            cursor = connections['sga_select'].cursor()
            sql = """SELECT  insti.codigo, insti.nombre , ca.codigo AS codigo_carrera ,'cedula' AS tipo_identificacion, perso.cedula AS identifiacion,
                                        (SELECT se.nombre
                                        FROM sga_sexo se
                                        WHERE se.id=perso.sexo_id AND se.status= TRUE) AS sexo,  perso.apellido1 AS primer_apellido, perso.apellido2 AS segundo_apellido, perso.nombres AS nombres,
                                        ins.fechainicioprimernivel AS fecha_inicio_primer_nivel, ins.fechainicioconvalidacion AS fecha_ingreso_convalidacion,
                                        (SELECT grad.fechagraduado FROM  sga_graduado grad
                                        WHERE grad.inscripcion_id=ins.id) AS fecha_graduacion, 
                                        '' AS fecha_inicio, '' AS fecha_fin,
                                        '' AS total_creditos_sin_titulacion, 
                                         (SELECT round(SUM(recordac.creditos),2)
                                        FROM sga_recordacademico recordac
                                        WHERE (recordac.inscripcion_id=ins.id AND recordac.status AND recordac.valida=True)) AS numero_creditos_aprobados,
                                        ins.id AS idinscripcion
                                        FROM sga_matricula matri, sga_nivel ni,sga_inscripcion ins, sga_persona perso,sga_periodo peri,sga_nivelmalla nimalla,sga_sesion sesion, sga_modalidad modal,
                                        sga_institucioneducacionsuperior insti, sga_carrera ca
                                        WHERE matri.nivel_id=ni.id AND ni.periodo_id=peri.id AND matri.nivelmalla_id=nimalla.id AND insti.id=1024
                                        AND matri.inscripcion_id=ins.id AND ins.persona_id=perso.id AND ins.sesion_id=sesion.id 
                                        AND ins.modalidad_id=modal.id AND ins.status= TRUE AND ni.periodo_id='%s' 
                                        AND matri.status= TRUE AND ni.status= TRUE AND matri.retiradomatricula= FALSE 
                                        AND ins.coordinacion_id =1 AND matri.nivelmalla_id=8 AND ins.carrera_id=1 AND ins.carrera_id=ca.id 
                                        union all
                                        SELECT  insti.codigo, insti.nombre , ca.codigo AS codigo_carrera ,'cedula' AS tipo_identificacion, perso.cedula AS identifiacion,
                                        (SELECT se.nombre
                                        FROM sga_sexo se
                                        WHERE se.id=perso.sexo_id AND se.status= TRUE) AS sexo,  perso.apellido1 AS primer_apellido, perso.apellido2 AS segundo_apellido, perso.nombres AS nombres,
                                        ins.fechainicioprimernivel AS fecha_inicio_primer_nivel, ins.fechainicioconvalidacion AS fecha_ingreso_convalidacion,
                                        (SELECT grad.fechagraduado FROM  sga_graduado grad
                                        WHERE grad.inscripcion_id=ins.id) AS fecha_graduacion, 
                                        '' AS fecha_inicio, '' AS fecha_fin,
                                        '' AS total_creditos_sin_titulacion, 
                                         (SELECT round(SUM(recordac.creditos),2)
                                        FROM sga_recordacademico recordac
                                        WHERE (recordac.inscripcion_id=ins.id AND recordac.status AND recordac.valida=True)) AS numero_creditos_aprobados,
                                        ins.id AS idinscripcion
                                        FROM sga_matricula matri, sga_nivel ni,sga_inscripcion ins, sga_persona perso,sga_periodo peri,sga_nivelmalla nimalla,sga_sesion sesion, sga_modalidad modal,
                                        sga_institucioneducacionsuperior insti, sga_carrera ca
                                        WHERE matri.nivel_id=ni.id AND ni.periodo_id=peri.id AND matri.nivelmalla_id=nimalla.id AND insti.id=1024
                                        AND matri.inscripcion_id=ins.id AND ins.persona_id=perso.id AND ins.sesion_id=sesion.id 
                                        AND ins.modalidad_id=modal.id AND ins.status= TRUE AND ni.periodo_id='%s' 
                                        AND matri.status= TRUE AND ni.status= TRUE AND matri.retiradomatricula= FALSE 
                                        AND ins.coordinacion_id =1 AND matri.nivelmalla_id=9 AND ins.carrera_id=110 AND ins.carrera_id=ca.id ORDER BY 7,8""" % (periodo, periodo)
            cursor.execute(sql)
            results = cursor.fetchall()
            row_num = 1
            for r in results:
                i = 0
                campo1 = u"%s" % r[0]
                campo2 = u"%s" % r[1]
                campo3 = u"%s" % r[2]
                campo4 = u"%s" % r[3]
                campo5 = u"%s" % r[4]
                campo6 = u"%s" % r[5]
                campo7 = u"%s" % r[6]
                campo8 = u"%s" % r[7]
                campo9 = u"%s" % r[8]
                campo10 = u"%s" % r[9]
                campo11 = u"%s" % r[10]
                campo12 = u"%s" % r[11]
                campo13 = u"%s" % r[12]
                campo14 = u"%s" % r[13]
                # campo15 = u"%s" % r[14]
                campo16 = u"%s" % r[15]
                inscripcion = Inscripcion.objects.get(pk=r[16])
                campo15 = inscripcion.total_creditos_malla()
                ws.write(row_num, 0, str(campo1), formatoceldaleft)
                ws.write(row_num, 1, str(campo2), formatoceldaleft)
                ws.write(row_num, 2, str(campo3), formatoceldaleft)
                ws.write(row_num, 3, str(campo4), formatoceldaleft)
                ws.write(row_num, 4, str(campo5), formatoceldaleft)
                ws.write(row_num, 5, str(campo6), formatoceldaleft)
                ws.write(row_num, 6, str(campo7), formatoceldaleft)
                ws.write(row_num, 7, str(campo8), formatoceldaleft)
                ws.write(row_num, 8, str(campo9), formatoceldaleft)
                ws.write(row_num, 9, str(campo10), formatoceldaleft)
                ws.write(row_num, 10, str(campo11), formatoceldaleft)
                ws.write(row_num, 11, str(campo12), formatoceldaleft)
                ws.write(row_num, 12, str(campo13), formatoceldaleft)
                ws.write(row_num, 13, str(campo14), formatoceldaleft)
                ws.write(row_num, 14, '', formatoceldaleft)
                ws.write(row_num, 15, str(campo15), formatoceldaleft)
                row_num += 1
            workbook.close()

            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.cuerpo = 'Excel listo'
                noti.url = "{}reportes/alumnos/{}".format(MEDIA_URL, nombre_archivo)
                noti.save()
            else:
                noti = Notificacion(cuerpo='Reporte listo',
                                    titulo='Reporte de aptos EHEP',
                                    destinatario=ePersona,
                                    url="{}reportes/alumnos/{}".format(MEDIA_URL, nombre_archivo),
                                    prioridad=1,
                                    app_label='SGA',
                                    fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2,
                                    en_proceso=False)
                noti.save(request)

            try:
                send_user_notification(user=user, payload={
                    "head": noti.titulo,
                    "body": noti.cuerpo,
                    "action": "notificacion",
                    "timestamp": time.mktime(datetime.now().timetuple()),
                    "url": noti.url,
                    "btn_notificaciones": traerNotificaciones(request, data, ePersona),
                    "mensaje": 'Su reporte ha sido generado con exito'
                }, ttl=500)
            except:
                pass

        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)
            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.cuerpo = 'Error en el reporte {} - linea {}'.format(ex, sys.exc_info()[-1].tb_lineno)
                noti.url = ""
                noti.save()
            else:
                noti = Notificacion(cuerpo='Error en el reporte {} - linea {}'.format(ex, sys.exc_info()[-1].tb_lineno),
                                    titulo=f'Reporte de aptos EHEP',
                                    destinatario=ePersona, url="",
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)
                noti.save(request)
            try:
                send_user_notification(user=user, payload={
                    "head": noti.titulo,
                    "body": noti.cuerpo,
                    "action": "notificacion",
                    "timestamp": time.mktime(datetime.now().timetuple()),
                    "url": noti.url,
                    "btn_notificaciones": traerNotificaciones(request, data, ePersona),
                    "mensaje": textoerror
                }, ttl=500)
            except:
                pass


class reporte_cumplimiento_titulacion(threading.Thread):

    def __init__(self, request, data, notiid, periodo):
        self.request = request
        self.data = data
        self.notiid = notiid
        self.periodo = periodo
        threading.Thread.__init__(self)

    def run(self):
        from sga.models import Matricula, Persona, MateriaAsignada, Periodo, Malla, RequisitoTitulacionMalla,\
            Notificacion
        try:
            directory = os.path.join(MEDIA_ROOT, 'reportes', 'matrices')
            request, data, notiid, periodo_id = self.request, self.data, self.notiid, self.periodo
            usernotify = User.objects.get(pk=request.user.pk)
            pers = Persona.objects.get(usuario=usernotify)
            os.makedirs(directory, exist_ok=True)
            nombre_archivo = 'reporte_cumplimiento_titulacion_' + random.randint(1, 10000).__str__() + '.xls'
            directory = os.path.join(MEDIA_ROOT, 'reportes', 'matrices', nombre_archivo)
            try:
                ePeriodo = Periodo.objects.get(pk=periodo_id)
            except ObjectDoesNotExist:
                raise NameError(u"No se encontro periodo académico")
            __author__ = 'Unemi'
            title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
            font_style = XFStyle()
            font_style.font.bold = True
            font_style2 = XFStyle()
            font_style2.font.bold = False
            wb = Workbook(encoding='utf-8')
            filtro_matricula = Q(nivel__periodo_id=ePeriodo.id) & Q(estado_matricula__in=[2, 3]) & Q(status=True) &\
                               Q(retiradomatricula=False) & Q(inscripcion__inscripcionmalla__status=True) & \
                               Q(inscripcion__inscripcionmalla__malla__asignaturamalla__validarequisitograduacion=True)
            eMatriculas = Matricula.objects.filter(filtro_matricula).order_by('inscripcion__carrera')
            eMallas = Malla.objects.filter(pk__in=eMatriculas.values_list('inscripcion__inscripcionmalla__malla__id', flat=True).distinct())
            if DEBUG:
                eMallas = eMallas[:1]
            for eMalla in eMallas:
                _eMatriculas = eMatriculas.filter(inscripcion__inscripcionmalla__malla=eMalla, materiaasignada__materia__asignaturamalla__validarequisitograduacion=True)
                ws = wb.add_sheet(f'{eMalla.id}')
                ws.write_merge(0, 0, 0, 24, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                ws.write_merge(1, 1, 0, 24, f'{eMalla.__str__()}', title)
                columns = [
                    # (u"MALLA", 29192),
                    (u"FACULTAD", 16698),
                    (u"CARRERA", 8957),
                    (u"MATERIA", 16698),
                    (u"NIVEL MATERIA", 8957),
                    (u"MODALIDAD", 8957),
                    (u"NIVEL MALLA", 4083),
                    (u"CÉDULA", 3570),
                    (u"ESTUDIANTE", 13045),
                    (u"CORREO INSTITUCIONAL", 10720),
                    (u"INSCRIPCION", 3131),
                    (u"MATRICULA", 3131),
                    (u"MATERIAASIGNADA", 3131),
                ]

                eRequisitoTitulacionMallas = RequisitoTitulacionMalla.objects.filter(malla=eMalla, status=True).order_by('requisito__nombre')
                aRequisitos = []
                num = 1
                for eRequisitoTitulacionMalla in eRequisitoTitulacionMallas:
                    num += 1
                    columns.append((f"R{num}", 7000, 0))
                    aRequisitos.append({'id': eRequisitoTitulacionMalla.id,
                                        'name': eRequisitoTitulacionMalla.requisito.nombre,
                                        'num': num})

                row_num = 2
                for col_num in range(len(columns)):
                    ws.write(row_num, col_num, columns[col_num][0], font_style)
                    ws.col(col_num).width = columns[col_num][1]
                total = _eMatriculas.count()
                row_num = 3
                cont = 1
                for eMatricula in _eMatriculas:
                    eInscripcion = eMatricula.inscripcion
                    eNivelMalla = eMatricula.nivelmalla
                    eCoordinacion = eMatricula.nivel.coordinacion()
                    eCarrera = eInscripcion.carrera
                    ePersona = eInscripcion.persona
                    eModalidad = eMalla.modalidad
                    identificacion = ePersona.identificacion()
                    nombres = ePersona.nombre_completo_inverso()
                    correoinst = ePersona.emailinst
                    celular = ePersona.telefono
                    eMateriaAsignadas = MateriaAsignada.objects.filter(status=True, matricula=eMatricula, materia__asignaturamalla__validarequisitograduacion=True)
                    for eMateriaAsignada in eMateriaAsignadas:
                        eAsignaturaMalla = eMateriaAsignada.materia.asignaturamalla
                        eMateria = eMateriaAsignada.materia
                        eAsignatura = eAsignaturaMalla.asignatura
                        print(f"({total}/{cont}) -> {ePersona.nombre_completo_inverso()} >> {eAsignatura.nombre}")
                        # ws.write(row_num, 0, u'%s' % eMalla.__str__(), font_style2)
                        ws.write(row_num, 0, u'%s' % eCoordinacion.__str__(), font_style2)
                        ws.write(row_num, 1, u'%s' % eCarrera.__str__(), font_style2)
                        ws.write(row_num, 2, u'%s' % eAsignatura.nombre, font_style2)
                        ws.write(row_num, 3, u'%s' % eAsignaturaMalla.nivelmalla.__str__(), font_style2)
                        ws.write(row_num, 4, u'%s' % eModalidad.__str__(), font_style2)
                        ws.write(row_num, 5, u'%s' % eNivelMalla.nombre, font_style2)
                        ws.write(row_num, 6, u'%s' % identificacion, font_style2)
                        ws.write(row_num, 7, u'%s' % nombres, font_style2)
                        ws.write(row_num, 8, u'%s' % correoinst, font_style2)
                        ws.write(row_num, 9, u'%s' % eInscripcion.id, font_style2)
                        ws.write(row_num, 10, u'%s' % eMatricula.id, font_style2)
                        ws.write(row_num, 11, u'%s' % eMateriaAsignada.id, font_style2)
                        col = 11
                        if len(eRequisitos := eMateria.requisitomateriaunidadintegracioncurricular_set.filter(status=True).order_by('requisito__nombre')) > 0:
                            for aRequisito in aRequisitos:
                                col += 1
                                idRequisito = aRequisito.get('id', '0')
                                eRequisito = eRequisitos.filter(requisito_id=idRequisito).first()
                                if eRequisito:
                                    cumple = eRequisito.run(eInscripcion.pk)
                                    ws.write(row_num, col, "%s" % 'SI' if cumple else 'NO', font_style2)  # REQUISITO
                                else:
                                    ws.write(row_num, col, "-", font_style2)  # REQUISITO
                        cont += 1
                        row_num += 1
                col += 2

                _col = col
                ws.write(2, _col, "DESCRIPCIÓN", font_style)
                ws.write(2, _col + 1, 'Nro', font_style)
                row_num = 3
                for aRequisito in aRequisitos:
                    name = aRequisito.get('name', None)
                    num = aRequisito.get('num', '0')
                    ws.write(row_num, _col, u'%s' % name, font_style)
                    ws.write(row_num, _col + 1, u'R%s' % num, font_style2)
                    row_num += 1
            wb.save(directory)
            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.cuerpo = 'Excel Listo'
                noti.url = "{}reportes/matrices/{}".format(MEDIA_URL, nombre_archivo)
                noti.save()
            else:
                noti = Notificacion(cuerpo='Reporte Listo',
                                    titulo='Reporte cumplimiento titulacion',
                                    destinatario=pers,
                                    url="{}reportes/matrices/{}".format(MEDIA_URL, nombre_archivo),
                                    prioridad=1,
                                    app_label='SGA',
                                    fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2,
                                    en_proceso=False)

                noti.save(request)

            try:
                send_user_notification(user=usernotify, payload={
                    "head": noti.titulo,
                    "body": noti.cuerpo,
                    "action": "notificacion",
                    "timestamp": time.mktime(datetime.now().timetuple()),
                    "url": noti.url,
                    "btn_notificaciones": traerNotificaciones(request, data, pers),
                    "mensaje": 'Su reporte ha sido generado con exito'
                }, ttl=500)
            except:
                pass
        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)
            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.cuerpo = 'Error en el reporte {} - linea {}'.format(ex, sys.exc_info()[-1].tb_lineno)
                noti.url = ""
                noti.save()
            else:
                noti = Notificacion(cuerpo='Error en el reporte {} - linea {}'.format(ex, sys.exc_info()[-1].tb_lineno),
                                    titulo=f'Reporte cumplimiento titulacion',
                                    destinatario=pers,
                                    url="",
                                    prioridad=1,
                                    app_label='SGA',
                                    fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2,
                                    en_proceso=False)
                noti.save(request)
            try:
                send_user_notification(user=usernotify, payload={
                    "head": noti.titulo,
                    "body": noti.cuerpo,
                    "action": "notificacion",
                    "timestamp": time.mktime(datetime.now().timetuple()),
                    "url": noti.url,
                    "btn_notificaciones": traerNotificaciones(request, data, pers),
                    "mensaje": textoerror
                }, ttl=500)
            except:
                pass


class reporte_alumnos_aprobado_reprobado_nivelacion(threading.Thread):

    def __init__(self, request, data, notiid, periodo):
        self.request = request
        self.data = data
        self.notiid = notiid
        self.periodo = periodo
        threading.Thread.__init__(self)

    def run(self):
        from sga.models import Matricula, Persona, MateriaAsignada, Periodo, Malla, RequisitoTitulacionMalla,\
            Notificacion
        try:
            directory = os.path.join(MEDIA_ROOT, 'reportes', 'alumnos')
            request, data, notiid, periodo_id = self.request, self.data, self.notiid, self.periodo
            usernotify = User.objects.get(pk=request.user.pk)
            pers = Persona.objects.get(usuario=usernotify)
            os.makedirs(directory, exist_ok=True)
            nombre_archivo = 'reporte_aprobado_reprobado_nivelacion_' + random.randint(1, 10000).__str__() + '.xls'
            directory = os.path.join(MEDIA_ROOT, 'reportes', 'alumnos', nombre_archivo)
            try:
                ePeriodo = Periodo.objects.get(pk=periodo_id)
            except ObjectDoesNotExist:
                raise NameError(u"No se encontro periodo académico")
            __author__ = 'Unemi'
            title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
            font_style = XFStyle()
            font_style.font.bold = True
            font_style2 = XFStyle()
            font_style2.font.bold = False
            wb = Workbook(encoding='utf-8')
            ws = wb.add_sheet('Hoja1')
            ws.write_merge(0, 0, 0, 7, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
            columns = [
                (u"FACULTAD", 6000),
                (u"CARRERA", 6000),
                (u"CEDULA", 3000),
                (u"PRIMER APELLIDO", 6000),
                (u"SEGUNDO APELLIDO", 6000),
                (u"NOMBRES", 6000),
                (u"EMAIL", 6000),
                (u"EMAIL INSTITUCIONAL", 6000),
                (u"CELULAR", 3000),
                (u"TELEFONO", 3000),
                (u"MATERIA", 6000),
                (u"PARALELO", 3000),
                (u"NIVEL", 3000),
                (u"NOTA FINAL", 6000),
                (u"ASISTENCIA FINAL", 6000),
                (u"ESTADO", 3000),
                (u"NUMERO DE MATRICULA", 3000),
                (u"CUPO", 3000)
            ]
            row_num = 3
            for col_num in range(len(columns)):
                ws.write(row_num, col_num, columns[col_num][0], font_style)
                ws.col(col_num).width = columns[col_num][1]
            cursor = connections['sga_select'].cursor()
            sql = """
                                        select coordinacion.nombre as coordinacion, CASE WHEN carrera.mencion = '' THEN carrera.nombre ELSE (carrera.nombre || ' CON MENCION EN ' || carrera.mencion) END as carrera, 
                                        persona.cedula, persona.apellido1, persona.apellido2, persona.nombres, persona.email, persona.emailinst, persona.telefono, persona.telefono_conv, 
                                        (asignatura.nombre || CASE WHEN materia.alias = '' THEN '' ELSE (' - ('|| materia.alias || ')') END  || CASE WHEN materia.identificacion = '' THEN '' ELSE (' - ['|| materia.identificacion || ']') END) as materia, 
                                        materia.paralelo as paralelo, nivelmalla.nombre as nivel, materiaasig.notafinal as notafinal, materiaasig.asistenciafinal as asistenciafinal, CASE WHEN materiaasig.notafinal >=70 THEN 'APROBADO' ELSE 'REPROBADO' END,  materiaasig.matriculas as veces, matricula.aprobado as cupo   
                                        from sga_materiaasignada materiaasig 
                                        inner join sga_materia materia on materia.id = materiaasig.materia_id 
                                        inner join sga_nivel nivel on nivel.id = materia.nivel_id 
                                        inner join sga_asignaturamalla asigmalla on asigmalla.id = materia.asignaturamalla_id 
                                        inner join sga_malla malla on malla.id = asigmalla.malla_id 
                                        inner join sga_carrera carrera on carrera.id = malla.carrera_id 
                                        inner join sga_matricula matricula on matricula.id = materiaasig.matricula_id 
                                        inner join sga_inscripcion inscripcion on inscripcion.id = matricula.inscripcion_id 
                                        inner join sga_persona persona on persona.id = inscripcion.persona_id 
                                        inner join sga_asignatura asignatura on asignatura.id = materia.asignatura_id 
                                        inner join sga_nivelmalla nivelmalla on nivelmalla.id = asigmalla.nivelmalla_id 
                                        inner join sga_tipoestado tipoestado on tipoestado.id = materiaasig.estado_id 
                                        inner join sga_coordinacion_carrera coordinacioncarrera on coordinacioncarrera.carrera_id = carrera.id 
                                        inner join sga_coordinacion coordinacion on coordinacion.id = coordinacioncarrera.coordinacion_id 
                                        where matricula.status=true and materiaasig.retiramateria=false and materiaasig.status=true and coordinacion.id in(9) and nivel.periodo_id = %s
                                """ % ePeriodo.id
            cursor.execute(sql)
            results = cursor.fetchall()
            row_num = 4
            for r in results:
                i = 0
                campo1 = r[0]
                campo2 = r[1]
                campo3 = r[2]
                campo4 = r[3]
                campo5 = r[4]
                campo6 = r[5]
                campo7 = r[6]
                campo8 = r[7]
                campo9 = r[8]
                campo10 = r[9]
                campo11 = r[10]
                campo12 = r[11]
                campo13 = r[12]
                campo14 = r[13]
                campo15 = r[14]
                campo16 = r[15]
                campo17 = r[16]
                campo18 = r[17]
                ws.write(row_num, 0, "%s" % campo1, font_style2)
                ws.write(row_num, 1, "%s" % campo2, font_style2)
                ws.write(row_num, 2, "%s" % campo3, font_style2)
                ws.write(row_num, 3, "%s" % campo4, font_style2)
                ws.write(row_num, 4, "%s" % campo5, font_style2)
                ws.write(row_num, 5, "%s" % campo6, font_style2)
                ws.write(row_num, 6, "%s" % campo7, font_style2)
                ws.write(row_num, 7, "%s" % campo8, font_style2)
                ws.write(row_num, 8, "%s" % campo9, font_style2)
                ws.write(row_num, 9, "%s" % campo10, font_style2)
                ws.write(row_num, 10, "%s" % campo11, font_style2)
                ws.write(row_num, 11, "%s" % campo12, font_style2)
                ws.write(row_num, 12, "%s" % campo13, font_style2)
                ws.write(row_num, 13, "%d" % campo14, font_style2)
                ws.write(row_num, 14, "%d" % campo15, font_style2)
                ws.write(row_num, 15, "%s" % campo16, font_style2)
                ws.write(row_num, 16, "%d" % campo17, font_style2)
                ws.write(row_num, 17, "%d" % campo18, font_style2)
                row_num += 1
            wb.save(directory)
            usernotify = User.objects.get(pk=request.user.pk)
            pers = Persona.objects.get(usuario=usernotify)
            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.cuerpo = 'Excel Listo'
                noti.url = "{}reportes/alumnos/{}".format(MEDIA_URL, nombre_archivo)
                noti.save()
            else:
                noti = Notificacion(cuerpo='Reporte Listo',
                                    titulo='Reporte de estudiantes aprobados y reprobados de nivelación',
                                    destinatario=pers,
                                    url="{}reportes/alumnos/{}".format(MEDIA_URL, nombre_archivo),
                                    prioridad=1,
                                    app_label='SGA',
                                    fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2,
                                    en_proceso=False)

                noti.save(request)

            try:
                send_user_notification(user=usernotify, payload={
                    "head": noti.titulo,
                    "body": noti.cuerpo,
                    "action": "notificacion",
                    "timestamp": time.mktime(datetime.now().timetuple()),
                    "url": noti.url,
                    "btn_notificaciones": traerNotificaciones(request, data, pers),
                    "mensaje": 'Su reporte ha sido generado con exito'
                }, ttl=500)
            except:
                pass
        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)
            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.cuerpo = 'Error en el reporte {} - linea {}'.format(ex, sys.exc_info()[-1].tb_lineno)
                noti.url = ""
                noti.save()
            else:
                noti = Notificacion(cuerpo='Error en el reporte {} - linea {}'.format(ex, sys.exc_info()[-1].tb_lineno),
                                    titulo=f'Reporte cumplimiento titulacion',
                                    destinatario=pers,
                                    url="",
                                    prioridad=1,
                                    app_label='SGA',
                                    fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2,
                                    en_proceso=False)
                noti.save(request)
            try:
                send_user_notification(user=usernotify, payload={
                    "head": noti.titulo,
                    "body": noti.cuerpo,
                    "action": "notificacion",
                    "timestamp": time.mktime(datetime.now().timetuple()),
                    "url": noti.url,
                    "btn_notificaciones": traerNotificaciones(request, data, pers),
                    "mensaje": textoerror
                }, ttl=500)
            except:
                pass


class reporte_avance_asistencia(threading.Thread):

    def __init__(self, request, data, notiid, periodo, fecha):
        self.request = request
        self.data = data
        self.notiid = notiid
        self.periodo = periodo
        self.fecha = fecha
        threading.Thread.__init__(self)

    def run(self):
        from sga.models import Matricula, Persona, MateriaAsignada, Periodo, Malla, RequisitoTitulacionMalla,\
            Notificacion, Coordinacion, DiasNoLaborable
        try:
            directory = os.path.join(MEDIA_ROOT, 'reportes', 'asistencias')
            request, data, notiid, periodo_id, fecha = self.request, self.data, self.notiid, self.periodo, self.fecha
            usernotify = User.objects.get(pk=request.user.pk)
            pers = Persona.objects.get(usuario=usernotify)
            os.makedirs(directory, exist_ok=True)
            nombre_archivo = 'reporte_avance_asistencia_' + random.randint(1, 10000).__str__() + '.xls'
            directory = os.path.join(MEDIA_ROOT, 'reportes', 'asistencias', nombre_archivo)
            try:
                ePeriodo = Periodo.objects.get(pk=periodo_id)
            except ObjectDoesNotExist:
                raise NameError(u"No se encontro periodo académico")
            eCoordinaciones = Coordinacion.objects.filter(carrera__malla__asignaturamalla__materia__nivel__periodo=ePeriodo).order_by('-id').distinct()
            diasnolaborables = DiasNoLaborable.objects.filter(periodo=ePeriodo).order_by('fecha')
            __author__ = 'Unemi'

            style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
            style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
            style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
            title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
            style1 = easyxf(num_format_str='D-MMM-YY')
            font_style = XFStyle()
            font_style.font.bold = True
            font_style2 = XFStyle()
            font_style2.font.bold = False
            wb = Workbook(encoding='utf-8')
            for eCoordinacion in eCoordinaciones:
                ws = wb.add_sheet(eCoordinacion.alias)
                ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                ws.write(1, 0, "Periodo: " + ePeriodo.nombre, font_style2)
                ws.write(2, 0, "Fecha Corte: " + request.POST['fecha'], font_style2)
                columns = [
                    (u"FACULTAD", 6000),
                    (u"CARRERA", 6000),
                    (u"NIVEL", 6000),
                    (u"PARALELO", 6000),
                    (u"DOCENTE", 6000),
                    (u"ASIGNATURA", 6000),
                    (u"HORAS PROGRAMADAS MENSUAL", 6000),
                    (u"HORAS INGRESADAS", 6000),
                    (u"HORAS FALTAS", 6000),
                    (u"HORAS TOTAL ASISTENCIAS", 6000),
                    (u"PORCENTAJE AVANCE FECHA TOPE", 6000),
                    (u"PORCENTAJE AVANCE SEMESTRE", 6000),
                    (u"FECHA INICIO MATERIA", 6000),
                    (u"FECHA FIN MATERIA", 6000),
                ]
                row_num = 3
                for col_num in range(len(columns)):
                    ws.write(row_num, col_num, columns[col_num][0], font_style)
                    ws.col(col_num).width = columns[col_num][1]
                cursor = connections['sga_select'].cursor()
                sql = "select tabladocente.facultad , tabladocente.carrera, tabladocente.nivel, tabladocente.paralelo, tabladocente.docente,tabladocente.asignatura, tabladocente.horas , COALESCE(tablaasistencia.asistencia,0) as asistencia, tabladocente.inicio, tabladocente.fin,  " \
                      " (select count(*) from sga_faltasmateriaperiodo fm1 where fm1.materia_id=tabladocente.materiaid) as faltas " \
                      " from (select pm.id as profesormateriaid, m.inicio, m.fin, m.id as materiaid, co.nombre as facultad, ca.nombre as carrera, nm.nombre as nivel, m.paralelo, (per.apellido1 || ' ' || per.apellido2 || ' ' || per.nombres) as docente,asi.nombre as asignatura, m.horas " \
                      " from sga_profesormateria pm, sga_materia m, sga_nivel n, sga_asignaturamalla am, sga_malla ma,sga_carrera ca, sga_coordinacion_carrera cc, sga_coordinacion co, sga_nivelmalla nm, sga_profesor pr, sga_persona per, sga_asignatura asi " \
                      " where m.id=pm.materia_id and pm.tipoprofesor_id not in (4) and n.id=m.nivel_id and am.id=m.asignaturamalla_id and ma.id=am.malla_id and ca.id=ma.carrera_id and cc.carrera_id=ca.id and co.id=cc.coordinacion_id and nm.id=am.nivelmalla_id and pr.id=pm.profesor_id and per.id=pr.persona_id " \
                      " and asi.id=m.asignatura_id and n.periodo_id= " + str(ePeriodo.id) + " and co.id=" + str(eCoordinacion.id) + " order by co.nombre, ca.nombre, nm.nombre, m.paralelo, docente) as tabladocente " \
                                      " left join (select mat1.id as materiaid, count(mat1.id) as asistencia from sga_leccion l1 , sga_clase c1 , sga_materia mat1, sga_nivel ni1 where l1.clase_id=c1.id and c1.materia_id=mat1.id and mat1.nivel_id=ni1.id and l1.fecha<= '" + str(fecha) + "' and ni1.periodo_id=" + str(ePeriodo.id) + " and l1.fecha not in (select dnl1.fecha from sga_diasnolaborable dnl1 where dnl1.periodo_id=" + str(ePeriodo.id) + ") GROUP by mat1.id) as tablaasistencia on tablaasistencia.materiaid=tabladocente.materiaid order by tabladocente.facultad , tabladocente.carrera, tabladocente.nivel, tabladocente.paralelo, tabladocente.docente"
                cursor.execute(sql)
                results = cursor.fetchall()
                row_num = 4
                for r in results:
                    i = 0
                    campo1 = r[0]
                    campo2 = r[1]
                    campo3 = r[2]
                    campo4 = r[3]
                    campo5 = r[4]
                    campo6 = r[5]
                    campo7 = r[6]
                    campo8 = r[7]
                    campo9 = 0
                    if r[6] > 0:
                        porcentajeperiodo = round((float(r[7]) / r[6]) * 100, 2)
                        if porcentajeperiodo > 100:
                            porcentajeperiodo = 100
                        campo9 = porcentajeperiodo
                    campo13 = 0
                    if (r[10] + r[7]) > 0:
                        porcentajefecha = round((float(r[7]) / (r[10] + r[7])) * 100, 2)
                        if porcentajefecha > 100:
                            porcentajefecha = 100
                        campo13 = porcentajefecha

                    campo10 = r[8]
                    campo11 = r[9]
                    campo12 = r[10]

                    ws.write(row_num, 0, campo1, font_style2)
                    ws.write(row_num, 1, campo2, font_style2)
                    ws.write(row_num, 2, campo3, font_style2)
                    ws.write(row_num, 3, campo4, font_style2)
                    ws.write(row_num, 4, campo5, font_style2)
                    ws.write(row_num, 5, campo6, font_style2)
                    ws.write(row_num, 6, campo7, font_style2)
                    ws.write(row_num, 7, campo8, font_style2)
                    ws.write(row_num, 8, campo12, font_style2)
                    ws.write(row_num, 9, campo12 + campo8, font_style2)
                    ws.write(row_num, 10, campo13, font_style2)
                    ws.write(row_num, 11, campo9, font_style2)
                    ws.write(row_num, 12, campo10, style1)
                    ws.write(row_num, 13, campo11, style1)
                    # while i < len(r):
                    #     # ws.write(row_num, i, r[i], font_style)
                    #     # ws.col(i).width = columns[i][1]
                    row_num += 1

            wb.save(directory)
            usernotify = User.objects.get(pk=request.user.pk)
            pers = Persona.objects.get(usuario=usernotify)
            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.cuerpo = 'Excel Listo'
                noti.url = "{}reportes/asistencias/{}".format(MEDIA_URL, nombre_archivo)
                noti.save()
            else:
                noti = Notificacion(cuerpo='Reporte Listo',
                                    titulo='Reporte de avance de asistencia',
                                    destinatario=pers,
                                    url="{}reportes/asistencias/{}".format(MEDIA_URL, nombre_archivo),
                                    prioridad=1,
                                    app_label='SGA',
                                    fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2,
                                    en_proceso=False)

                noti.save(request)

            try:
                send_user_notification(user=usernotify, payload={
                    "head": noti.titulo,
                    "body": noti.cuerpo,
                    "action": "notificacion",
                    "timestamp": time.mktime(datetime.now().timetuple()),
                    "url": noti.url,
                    "btn_notificaciones": traerNotificaciones(request, data, pers),
                    "mensaje": 'Su reporte ha sido generado con exito'
                }, ttl=500)
            except:
                pass
        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)
            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.cuerpo = 'Error en el reporte {} - linea {}'.format(ex, sys.exc_info()[-1].tb_lineno)
                noti.url = ""
                noti.save()
            else:
                noti = Notificacion(cuerpo='Error en el reporte {} - linea {}'.format(ex, sys.exc_info()[-1].tb_lineno),
                                    titulo=f'Reporte de avance de asistencia',
                                    destinatario=pers,
                                    url="",
                                    prioridad=1,
                                    app_label='SGA',
                                    fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2,
                                    en_proceso=False)
                noti.save(request)
            try:
                send_user_notification(user=usernotify, payload={
                    "head": noti.titulo,
                    "body": noti.cuerpo,
                    "action": "notificacion",
                    "timestamp": time.mktime(datetime.now().timetuple()),
                    "url": noti.url,
                    "btn_notificaciones": traerNotificaciones(request, data, pers),
                    "mensaje": textoerror
                }, ttl=500)
            except:
                pass



class reporte_estudiantes_modulo_ingles(threading.Thread):
    def __init__(self, request, data, notiid, periodo):
        self.request = request
        self.data = data
        self.notiid = notiid
        self.periodo = periodo
        threading.Thread.__init__(self)

    def run(self):
        from sga.models import Persona, Notificacion, MateriaAsignada, Coordinacion
        try:
            directory = os.path.join(MEDIA_ROOT, 'reportes', 'alumnos')
            request, data, notiid, periodo = self.request, self.data, self.notiid, self.periodo
            user = User.objects.get(pk=request.user.pk)
            ePersona = Persona.objects.get(usuario=user)
            os.makedirs(directory, exist_ok=True)
            nombre_archivo = "reporte_alumnos_matriculados_modulos_ingles_{}.xlsx".format(random.randint(1, 10000).__str__())
            directory = os.path.join(MEDIA_ROOT, 'reportes', 'alumnos', nombre_archivo)
            workbook = xlsxwriter.Workbook(directory, {'constant_memory': True})
            ws = workbook.add_worksheet('listado')
            titulo = workbook.add_format(
                {'align': 'center', 'valign': 'vcenter', 'font_name': 'Times New Roman', 'bold': True,
                 'font_color': 'blue', 'font_size': 18})
            font_style = workbook.add_format({'bold': True, 'border': 1, 'align': 'center', 'valign': 'vcenter'})
            font_style2 = workbook.add_format({'font_name': 'Arial', 'font_size': 10, 'bold': False})
            ws.merge_range(0, 0, 0, 7, 'UNIVERSIDAD ESTATAL DE MILAGRO', titulo)
            columns = [
                (u"FACULTAD", 40),
                (u"ID_MATERIA_ASIGNADA", 40),
                (u"IDENTIFICACIÓN", 40),
                (u"ESTUDIANTE", 40),
                (u"CARRERA", 40),
                (u"MODULO_INGLES", 40),  # Nueva columna para Coordinación Ingles
                (u"ASIGNATURA", 40),
                (u"PARALELO", 40),
                (u"NIVEL SEMESTRE", 40),
                (u"CORREO ELECTRÓNICO", 40),
                (u"CORREO INSTITUCIONAL", 40),
                (u"TELÉFONOS", 40),
                (u"ID_MATERIA", 40),
                (u"FECHA_MATRICULA", 40),
                (u"PPL", 40),
                (u"¿TIENE DISCAPACIDAD?", 40),
                (u"TIPO DISCAPACIDAD", 40),
            ]

            row_num = 3
            for col_num in range(len(columns)):
                ws.write(row_num, col_num, columns[col_num][0], font_style)
                ws.set_column(col_num, col_num, columns[col_num][1])
            row_num = 4
            eListadoEstudiantes = MateriaAsignada.objects.filter(
                status=True,
                matricula__status=True,
                retiramateria=False,
                matricula__nivel__periodo__id=periodo,
                materia__asignaturamalla__malla__carrera__coordinacion=6,
                # matricula__inscripcion__persona__id=315097
            ).order_by(
                'matricula__inscripcion_carrera_id',
                'matricula__inscripcionpersona_apellido1'
            ).distinct()

            for e in eListadoEstudiantes:
                if e.matricula:
                    estudinate = e.matricula.inscripcion.persona
                    carrera = e.matricula.inscripcion.carrera
                    coordinacion = Coordinacion.objects.filter(carrera=carrera).first()
                    facultad = coordinacion.nombre if coordinacion else 'N/A'
                    ws.write(row_num, 0, facultad, font_style2)
                    ws.write(row_num, 1, '%s' % e.id, font_style2)
                    ws.write(row_num, 2, '%s' % estudinate.identificacion(), font_style2)
                    ws.write(row_num, 3, '%s' % estudinate.nombre_completo_inverso(), font_style2)
                    ws.write(row_num, 4, '%s' % e.matricula.inscripcion.carrera, font_style2)
                    # Escribir las coordenaciones en columnas separadas
                    coordinaciones = e.materia.asignaturamalla.malla.carrera.coordinacion_carrera()
                    ws.write(row_num, 5, '%s' % coordinaciones.nombre, font_style2)  # Coordinación 1
                    ws.write(row_num, 6, '%s' % e.materia.asignatura.nombre, font_style2)
                    ws.write(row_num, 7, '%s' % e.materia.paralelo, font_style2)

                    _nivel = str(e.matricula.nivelmalla) if e.matricula else 'No matriculado en el periodo vigente',
                    ws.write(row_num, 8, '%s' % _nivel, font_style2)
                    ws.write(row_num, 9, '%s' % estudinate.email, font_style2)
                    ws.write(row_num, 10, '%s' % estudinate.emailinst, font_style2)
                    ws.write(row_num, 11, '%s' % estudinate.telefonos(), font_style2)
                    ws.write(row_num, 12, '%s' % e.materia.id, font_style2)
                    ws.write(row_num, 13, '%s' % e.fecha_creacion, font_style2)

                    ws.write(row_num, 14, '%s' % 'si' if estudinate.ppl else 'no', font_style2)
                    tienediscapacidad = 'NO'
                    tipodiscapacidad = 'NINGUNA'

                    if estudinate.tiene_discapasidad():
                        tienediscapacidad = 'SI'
                        if estudinate.tiene_discapasidad().filter(tipodiscapacidad__isnull=False).exists():
                            tipodiscapacidad = estudinate.tiene_discapasidad().first().tipodiscapacidad.nombre
                        else:
                            tipodiscapacidad = 'NO DETERMINADA'

                    ws.write(row_num, 15, '%s' % tienediscapacidad, font_style2)
                    ws.write(row_num, 16, '%s' % tipodiscapacidad, font_style2)

                    row_num += 1

            workbook.close()

            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.cuerpo = 'Excel listo'
                noti.url = "{}reportes/alumnos/{}".format(MEDIA_URL, nombre_archivo)
                noti.save()
            else:
                noti = Notificacion(cuerpo='Reporte listo',
                                    titulo='Reporte de módulo inglés',
                                    destinatario=ePersona,
                                    url="{}reportes/alumnos/{}".format(MEDIA_URL, nombre_archivo),
                                    prioridad=1,
                                    app_label='SGA',
                                    fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2,
                                    en_proceso=False)
                noti.save(request)

            try:
                send_user_notification(user=user, payload={
                    "head": noti.titulo,
                    "body": noti.cuerpo,
                    "action": "notificacion",
                    "timestamp": time.mktime(datetime.now().timetuple()),
                    "url": noti.url,
                    "btn_notificaciones": traerNotificaciones(request, data, ePersona),
                    "mensaje": 'Su reporte ha sido generado con éxito'
                }, ttl=500)
            except:
                pass
        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)
            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.cuerpo = 'Error en el reporte {} - linea {}'.format(ex, sys.exc_info()[-1].tb_lineno)
                noti.url = ""
                noti.save()
            else:
                noti = Notificacion(cuerpo='Error en el reporte {} - linea {}'.format(ex, sys.exc_info()[-1].tb_lineno),
                                    titulo=f'Reporte de módulo inglés',
                                    destinatario=ePersona, url="",
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)
                noti.save(request)
            try:
                send_user_notification(user=user, payload={
                    "head": noti.titulo,
                    "body": noti.cuerpo,
                    "action": "notificacion",
                    "timestamp": time.mktime(datetime.now().timetuple()),
                    "url": noti.url,
                    "btn_notificaciones": traerNotificaciones(request, data, ePersona),
                    "mensaje": textoerror
                }, ttl=500)
            except:
                pass


class reporte_estudiantes_modulo_informatica(threading.Thread):
    def __init__(self, request, data, notiid, periodo):
        self.request = request
        self.data = data
        self.notiid = notiid
        self.periodo = periodo
        threading.Thread.__init__(self)

    def run(self):
        from sga.models import Persona, Notificacion, MateriaAsignada, Coordinacion
        try:
            directory = os.path.join(MEDIA_ROOT, 'reportes', 'alumnos')
            request, data, notiid, periodo = self.request, self.data, self.notiid, self.periodo
            user = User.objects.get(pk=request.user.pk)
            ePersona = Persona.objects.get(usuario=user)
            os.makedirs(directory, exist_ok=True)
            nombre_archivo = "reporte_alumnos_matriculados_modulos_informatica_{}.xlsx".format(random.randint(1, 10000).__str__())
            directory = os.path.join(MEDIA_ROOT, 'reportes', 'alumnos', nombre_archivo)
            workbook = xlsxwriter.Workbook(directory, {'constant_memory': True})
            ws = workbook.add_worksheet('listado')
            titulo = workbook.add_format(
                {'align': 'center', 'valign': 'vcenter', 'font_name': 'Times New Roman', 'bold': True,
                 'font_color': 'blue', 'font_size': 18})
            font_style = workbook.add_format({'bold': True, 'border': 1, 'align': 'center', 'valign': 'vcenter'})
            font_style2 = workbook.add_format({'font_name': 'Arial', 'font_size': 10, 'bold': False})
            ws.merge_range(0, 0, 0, 7, 'UNIVERSIDAD ESTATAL DE MILAGRO', titulo)
            columns = [
                (u"FACULTAD", 40),
                (u"ID_MATERIA_ASIGNADA", 40),
                (u"IDENTIFICACIÓN", 40),
                (u"ESTUDIANTE", 40),
                (u"CARRERA", 40),
                (u"MODULO_INFORMATICA", 40),  # Nueva columna para Coordinación Computacion
                (u"ASIGNATURA", 40),
                (u"PARALELO", 40),
                (u"NIVEL SEMESTRE", 40),
                (u"CORREO ELECTRÓNICO", 40),
                (u"CORREO INSTITUCIONAL", 40),
                (u"TELÉFONOS", 40),
                (u"ID_MATERIA", 40),
                (u"FECHA_MATRICULA", 40),
                (u"PPL", 40),
                (u"¿TIENE DISCAPACIDAD?", 40),
                (u"TIPO DISCAPACIDAD", 40),
            ]

            row_num = 3
            for col_num in range(len(columns)):
                ws.write(row_num, col_num, columns[col_num][0], font_style)
                ws.set_column(col_num, col_num, columns[col_num][1])
            row_num = 4
            eListadoEstudiantes = MateriaAsignada.objects.filter(
                status=True,
                matricula__status=True,
                retiramateria=False,
                matricula__nivel__periodo__id=periodo,
                materia__asignaturamalla__malla__carrera__coordinacion=8,
                 #matricula__inscripcion__persona__id=315097
            ).order_by(
                'matricula__inscripcion_carrera_id',
                'matricula__inscripcionpersona_apellido1'
            ).distinct()

            for e in eListadoEstudiantes:
                if e.matricula:
                    estudinate = e.matricula.inscripcion.persona
                    carrera = e.matricula.inscripcion.carrera
                    coordinacion = Coordinacion.objects.filter(carrera=carrera).first()
                    facultad = coordinacion.nombre if coordinacion else 'N/A'
                    ws.write(row_num, 0, facultad, font_style2)
                    ws.write(row_num, 1, '%s' % e.id, font_style2)
                    ws.write(row_num, 2, '%s' % estudinate.identificacion(), font_style2)
                    ws.write(row_num, 3, '%s' % estudinate.nombre_completo_inverso(), font_style2)
                    ws.write(row_num, 4, '%s' % e.matricula.inscripcion.carrera, font_style2)
                    # Escribir las coordenaciones en columnas separadas
                    coordinaciones = e.materia.asignaturamalla.malla.carrera.coordinacion_carrera()
                    ws.write(row_num, 5, '%s' % coordinaciones.nombre, font_style2)  # Coordinación 1
                    ws.write(row_num, 6, '%s' % e.materia.asignatura.nombre, font_style2)
                    ws.write(row_num, 7, '%s' % e.materia.paralelo, font_style2)

                    _nivel = str(e.matricula.nivelmalla) if e.matricula else 'No matriculado en el periodo vigente',
                    ws.write(row_num, 8, '%s' % _nivel, font_style2)
                    ws.write(row_num, 9, '%s' % estudinate.email, font_style2)
                    ws.write(row_num, 10, '%s' % estudinate.emailinst, font_style2)
                    ws.write(row_num, 11, '%s' % estudinate.telefonos(), font_style2)
                    ws.write(row_num, 12, '%s' % e.materia.id, font_style2)
                    ws.write(row_num, 13, '%s' % e.fecha_creacion, font_style2)

                    ws.write(row_num, 14, '%s' % 'si' if estudinate.ppl else 'no', font_style2)
                    tienediscapacidad = 'NO'
                    tipodiscapacidad = 'NINGUNA'

                    if estudinate.tiene_discapasidad():
                        tienediscapacidad = 'SI'
                        if estudinate.tiene_discapasidad().filter(tipodiscapacidad__isnull=False).exists():
                            tipodiscapacidad = estudinate.tiene_discapasidad().first().tipodiscapacidad.nombre
                        else:
                            tipodiscapacidad = 'NO DETERMINADA'

                    ws.write(row_num, 15, '%s' % tienediscapacidad, font_style2)
                    ws.write(row_num, 16, '%s' % tipodiscapacidad, font_style2)

                    row_num += 1

            workbook.close()

            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.cuerpo = 'Excel listo'
                noti.url = "{}reportes/alumnos/{}".format(MEDIA_URL, nombre_archivo)
                noti.save()
            else:
                noti = Notificacion(cuerpo='Reporte listo',
                                    titulo='Reporte de módulo informática',
                                    destinatario=ePersona,
                                    url="{}reportes/alumnos/{}".format(MEDIA_URL, nombre_archivo),
                                    prioridad=1,
                                    app_label='SGA',
                                    fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2,
                                    en_proceso=False)
                noti.save(request)

            try:
                send_user_notification(user=user, payload={
                    "head": noti.titulo,
                    "body": noti.cuerpo,
                    "action": "notificacion",
                    "timestamp": time.mktime(datetime.now().timetuple()),
                    "url": noti.url,
                    "btn_notificaciones": traerNotificaciones(request, data, ePersona),
                    "mensaje": 'Su reporte ha sido generado con éxito'
                }, ttl=500)
            except:
                pass
        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)
            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.cuerpo = 'Error en el reporte {} - linea {}'.format(ex, sys.exc_info()[-1].tb_lineno)
                noti.url = ""
                noti.save()
            else:
                noti = Notificacion(cuerpo='Error en el reporte {} - linea {}'.format(ex, sys.exc_info()[-1].tb_lineno),
                                    titulo=f'Reporte de módulo informática',
                                    destinatario=ePersona, url="",
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)
                noti.save(request)
            try:
                send_user_notification(user=user, payload={
                    "head": noti.titulo,
                    "body": noti.cuerpo,
                    "action": "notificacion",
                    "timestamp": time.mktime(datetime.now().timetuple()),
                    "url": noti.url,
                    "btn_notificaciones": traerNotificaciones(request, data, ePersona),
                    "mensaje": textoerror
                }, ttl=500)
            except:
                pass


class reporte_estudiantes_ingles_niveles(threading.Thread):
    def __init__(self, request, data, notiid, periodo):
        self.request = request
        self.data = data
        self.notiid = notiid
        self.periodo = periodo
        threading.Thread.__init__(self)

    def run(self):
        from sga.models import Persona, Notificacion, MateriaAsignada, Coordinacion,RecordAcademico, Matricula, Malla
        try:
            directory = os.path.join(MEDIA_ROOT, 'reportes', 'alumnos')
            request, data, notiid, periodo = self.request, self.data, self.notiid, self.periodo
            user = User.objects.get(pk=request.user.pk)
            ePersona = Persona.objects.get(usuario=user)
            os.makedirs(directory, exist_ok=True)
            nombre_archivo = "reporte_alumnos_aprobados_niveles_{}.xlsx".format(random.randint(1, 10000).__str__())
            directory = os.path.join(MEDIA_ROOT, 'reportes', 'alumnos', nombre_archivo)
            workbook = xlsxwriter.Workbook(directory, {'constant_memory': True})
            ws = workbook.add_worksheet('listado')
            titulo = workbook.add_format(
                {'align': 'center', 'valign': 'vcenter', 'font_name': 'Times New Roman', 'bold': True,
                 'font_color': 'blue', 'font_size': 18})
            font_style = workbook.add_format({'bold': True, 'border': 1, 'align': 'center', 'valign': 'vcenter'})
            font_style2 = workbook.add_format({'font_name': 'Arial', 'font_size': 10, 'bold': False})
            ws.merge_range(0, 0, 0, 7, 'UNIVERSIDAD ESTATAL DE MILAGRO', titulo)
            columns = [
                (u"FACULTAD", 40),
                (u"IDENTIFICACIÓN", 40),
                (u"ESTUDIANTE", 40),
                (u"CARRERA", 40),
                (u"NIVEL SEMESTRE", 40),
                (u"CORREO ELECTRÓNICO", 40),
                (u"CORREO INSTITUCIONAL", 40),
                (u"TELÉFONOS", 40),
                (u"TIPO (B1 o B2)", 40),
                (u"APROBADO", 40)
            ]

            row_num = 3
            for col_num in range(len(columns)):
                ws.write(row_num, col_num, columns[col_num][0], font_style)
                ws.set_column(col_num, col_num, columns[col_num][1])
            row_num = 4
            eListadoEstudiantes = RecordAcademico.objects.filter(
                inscripcion__modalidad_id__in=[1, 3],
                status=True,
                inscripcion__status=True,
                aprobada=True,
                # inscripcion__persona__id=51637,
                inscripcion__matricula__nivel__periodo_id=periodo,
                asignatura_id__in=[783, 784, 785, 786, 1690, 1691, 2070]
            ).order_by(
                'inscripcion__persona_id',
                'inscripcion__carrera_id',
                'inscripcion__persona__apellido1'
            ).distinct('inscripcion__persona_id')
            for e in eListadoEstudiantes:
                if e.inscripcion:
                    ins = e.inscripcion.id
                    estudiante = e.inscripcion.persona
                    carrera = e.inscripcion.carrera
                    coordinacion = Coordinacion.objects.filter(carrera=carrera).first()
                    facultad = coordinacion.nombre if coordinacion else 'N/A'
                    matricula = Matricula.objects.filter(inscripcion=ins).first()
                    malla = Malla.objects.filter(carrera_id=carrera.id, vigente=True).first()
                    nivel_suficiencia = malla.nivelsuficiencia_id if malla else None

                    modalidad = e.inscripcion.modalidad_id
                    asignaturas_aprobadas = RecordAcademico.objects.filter(
                        inscripcion__persona=estudiante,
                        asignatura_id__in=[783, 784, 785, 786, 1690, 1691, 2070],
                        aprobada=True
                    ).distinct('asignatura_id').count()

                    tipo = 'No cumple'
                    if e.inscripcion.modalidad_id == 1:
                        if nivel_suficiencia == 3 and asignaturas_aprobadas >= 5:
                            tipo = 'B1'
                        elif asignaturas_aprobadas >= 5 and nivel_suficiencia is  None:
                            tipo = 'B1'
                    elif e.inscripcion.modalidad_id == 3:
                        if nivel_suficiencia == 4 and asignaturas_aprobadas >= 7:
                            tipo = 'B2'
                        elif asignaturas_aprobadas >= 7 and nivel_suficiencia is  None:
                            tipo = 'B2'
                        elif nivel_suficiencia == 3 and asignaturas_aprobadas >= 5:
                            tipo = 'B1'

                    nivelmalla = matricula.nivelmalla if matricula else 'No matriculado en el periodo vigente'


                    if tipo != 'No cumple':
                        ws.write(row_num, 0, facultad, font_style2)
                        ws.write(row_num, 1, '%s' % estudiante.identificacion(), font_style2)
                        ws.write(row_num, 2, '%s' % estudiante.nombre_completo_inverso(), font_style2)
                        ws.write(row_num, 3, '%s' % carrera, font_style2)
                        ws.write(row_num, 4, '%s' % nivelmalla, font_style2)
                        ws.write(row_num, 5, '%s' % estudiante.email, font_style2)
                        ws.write(row_num, 6, '%s' % estudiante.emailinst,font_style2)
                        ws.write(row_num, 7, '%s' % estudiante.telefono, font_style2)
                        ws.write(row_num, 8, '%s' % tipo, font_style2)
                        ws.write(row_num, 9, 'Sí', font_style2)

                        row_num += 1

            workbook.close()

            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.cuerpo = 'Excel listo'
                noti.url = "{}reportes/alumnos/{}".format(MEDIA_URL, nombre_archivo)
                noti.save()
            else:
                noti = Notificacion(cuerpo='Reporte listo',
                                    titulo='Reporte del módulo de inglés por niveles',
                                    destinatario=ePersona,
                                    url="{}reportes/alumnos/{}".format(MEDIA_URL, nombre_archivo),
                                    prioridad=1,
                                    app_label='SGA',
                                    fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2,
                                    en_proceso=False)
                noti.save(request)

            try:
                send_user_notification(user=user, payload={
                    "head": noti.titulo,
                    "body": noti.cuerpo,
                    "action": "notificacion",
                    "timestamp": time.mktime(datetime.now().timetuple()),
                    "url": noti.url,
                    "btn_notificaciones": traerNotificaciones(request, data, ePersona),
                    "mensaje": 'Su reporte ha sido generado con éxito'
                }, ttl=500)
            except:
                pass
        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)
            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.cuerpo = 'Error en el reporte {} - linea {}'.format(ex, sys.exc_info()[-1].tb_lineno)
                noti.url = ""
                noti.save()
            else:
                noti = Notificacion(cuerpo='Error en el reporte {} - linea {}'.format(ex, sys.exc_info()[-1].tb_lineno),
                                    titulo=f'Reporte del módulo de inglés por niveles',
                                    destinatario=ePersona, url="",
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)
                noti.save(request)
            try:
                send_user_notification(user=user, payload={
                    "head": noti.titulo,
                    "body": noti.cuerpo,
                    "action": "notificacion",
                    "timestamp": time.mktime(datetime.now().timetuple()),
                    "url": noti.url,
                    "btn_notificaciones": traerNotificaciones(request, data, ePersona),
                    "mensaje": textoerror
                }, ttl=500)
            except:
                pass


class reporte_estudiantes_informatico_modulo_1y2(threading.Thread):
    def __init__(self, request, data, notiid, periodo):
        self.request = request
        self.data = data
        self.notiid = notiid
        self.periodo = periodo
        threading.Thread.__init__(self)

    def run(self):
        from sga.models import Persona, Notificacion, Coordinacion,RecordAcademico, Matricula
        try:
            directory = os.path.join(MEDIA_ROOT, 'reportes', 'alumnos')
            request, data, notiid, periodo = self.request, self.data, self.notiid, self.periodo
            user = User.objects.get(pk=request.user.pk)
            ePersona = Persona.objects.get(usuario=user)
            os.makedirs(directory, exist_ok=True)
            nombre_archivo = "reporte_alumnos_aprobados_informatico_{}.xlsx".format(random.randint(1, 10000).__str__())
            directory = os.path.join(MEDIA_ROOT, 'reportes', 'alumnos', nombre_archivo)
            workbook = xlsxwriter.Workbook(directory, {'constant_memory': True})
            ws = workbook.add_worksheet('listado')
            titulo = workbook.add_format(
                {'align': 'center', 'valign': 'vcenter', 'font_name': 'Times New Roman', 'bold': True,
                 'font_color': 'blue', 'font_size': 18})
            font_style = workbook.add_format({'bold': True, 'border': 1, 'align': 'center', 'valign': 'vcenter'})
            font_style2 = workbook.add_format({'font_name': 'Arial', 'font_size': 10, 'bold': False})
            ws.merge_range(0, 0, 0, 7, 'UNIVERSIDAD ESTATAL DE MILAGRO', titulo)
            columns = [
                (u"FACULTAD", 40),
                (u"IDENTIFICACIÓN", 40),
                (u"ESTUDIANTE", 40),
                (u"CARRERA", 40),
                (u"MODULO DE COMPUTACION I WINDOWS - WORD - INTERNET", 40),
                (u"MODULO DE COMPUTACION II EXCEL - POWERPOINT", 40),
                (u"NIVEL SEMESTRE", 40),
                (u"CORREO ELECTRÓNICO", 40),
                (u"CORREO INSTITUCIONAL", 40),
                (u"TELÉFONOS", 40)
            ]

            row_num = 3
            for col_num in range(len(columns)):
                ws.write(row_num, col_num, columns[col_num][0], font_style)
                ws.set_column(col_num, col_num, columns[col_num][1])
            row_num = 4

            eListadoEstudiantes = RecordAcademico.objects.filter(
                status=True,
                inscripcion__status=True,
                aprobada=True,
                inscripcion__matricula__nivel__periodo_id=periodo,
                asignatura_id__in=[1053, 1054]
            ).order_by(
                'inscripcion__persona_id',
                'inscripcion__carrera_id',
                'inscripcion__persona__apellido1'
            ).distinct('inscripcion__persona_id')

            for e in eListadoEstudiantes:
                if e.inscripcion:
                    estudiante = e.inscripcion.persona
                    ins = e.inscripcion.id
                    carrera = e.inscripcion.carrera
                    matricula = Matricula.objects.filter(inscripcion=ins).first()
                    coordinacion = Coordinacion.objects.filter(carrera=carrera).first()
                    facultad = coordinacion.nombre if coordinacion else 'N/A'

                    nota_informatica_1 = RecordAcademico.objects.filter(
                        inscripcion__persona=estudiante,
                        asignatura_id=1053,
                        aprobada=True
                    ).first()

                    nota_informatica_2 = RecordAcademico.objects.filter(
                        inscripcion__persona=estudiante,
                        asignatura_id=1054,
                        aprobada=True
                    ).first()

                    nivelmalla = matricula.nivelmalla if matricula else 'No matriculado en el periodo vigente'
                    ws.write(row_num, 0, facultad, font_style2)
                    ws.write(row_num, 1, '%s' % estudiante.identificacion(), font_style2)
                    ws.write(row_num, 2, '%s' % estudiante.nombre_completo_inverso(), font_style2)
                    ws.write(row_num, 3, '%s' % carrera, font_style2)
                    ws.write(row_num, 4, '%s' % (nota_informatica_1.nota if nota_informatica_1 else 'No ha cursado'),font_style2)
                    ws.write(row_num, 5, '%s' % (nota_informatica_2.nota if nota_informatica_2 else 'No ha cursado'),font_style2)
                    ws.write(row_num, 6, '%s' % nivelmalla, font_style2)
                    ws.write(row_num, 7, '%s' % estudiante.email, font_style2)
                    ws.write(row_num, 8, '%s' % estudiante.emailinst, font_style2)
                    ws.write(row_num, 9, '%s' % estudiante.telefono, font_style2)

                    row_num += 1

            workbook.close()

            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.cuerpo = 'Excel listo'
                noti.url = "{}reportes/alumnos/{}".format(MEDIA_URL, nombre_archivo)
                noti.save()
            else:
                noti = Notificacion(cuerpo='Reporte listo',
                                    titulo='Reporte de módulo informática aprobado 1 y 2',
                                    destinatario=ePersona,
                                    url="{}reportes/alumnos/{}".format(MEDIA_URL, nombre_archivo),
                                    prioridad=1,
                                    app_label='SGA',
                                    fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2,
                                    en_proceso=False)
                noti.save(request)

            try:
                send_user_notification(user=user, payload={
                    "head": noti.titulo,
                    "body": noti.cuerpo,
                    "action": "notificacion",
                    "timestamp": time.mktime(datetime.now().timetuple()),
                    "url": noti.url,
                    "btn_notificaciones": traerNotificaciones(request, data, ePersona),
                    "mensaje": 'Su reporte ha sido generado con éxito'
                }, ttl=500)
            except:
                pass
        except Exception as ex:
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)
            if notiid > 0:
                noti = Notificacion.objects.get(pk=notiid)
                noti.en_proceso = False
                noti.cuerpo = 'Error en el reporte {} - linea {}'.format(ex, sys.exc_info()[-1].tb_lineno)
                noti.url = ""
                noti.save()
            else:
                noti = Notificacion(cuerpo='Error en el reporte {} - linea {}'.format(ex, sys.exc_info()[-1].tb_lineno),
                                    titulo=f'Reporte de módulo informática aprobado 1 y 2',
                                    destinatario=ePersona, url="",
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)
                noti.save(request)
            try:
                send_user_notification(user=user, payload={
                    "head": noti.titulo,
                    "body": noti.cuerpo,
                    "action": "notificacion",
                    "timestamp": time.mktime(datetime.now().timetuple()),
                    "url": noti.url,
                    "btn_notificaciones": traerNotificaciones(request, data, ePersona),
                    "mensaje": textoerror
                }, ttl=500)
            except:
                pass
