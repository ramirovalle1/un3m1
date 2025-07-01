# -*- coding: latin-1 -*-
import os

import xlwt
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models.query_utils import Q
from django.http import HttpResponseRedirect, JsonResponse
from django.shortcuts import render
from googletrans import Translator
from openpyxl import load_workbook

from decorators import secure_module, last_access
from settings import MEDIA_ROOT
from sga.commonviews import adduserdata
from sga.forms import ArchivoNotasForm, ArchivoNotasSistemaForm
from sga.funciones import log, MiPaginador, generar_nombre, convertirfecha2, null_to_decimal
from sga.models import ArchivoNotas, Persona, Inscripcion, AsignaturaMalla, RecordAcademico, ModuloMalla, \
    HistoricoRecordAcademico


@login_required(redirect_field_name='ret', login_url='/loginsga')
@secure_module
@last_access
@transaction.atomic()
def view(request):
    data = {}
    adduserdata(request, data)
    persona = request.session['persona']
    permiso = False
    if persona.usuario.is_superuser:
        permiso = True
    usuario_ingreso = persona.usuario
    if request.method == 'POST':
        action = request.POST['action']

        if action == 'add':
            try:
                f = ArchivoNotasForm(request.POST, request.FILES)
                if 'archivo' in request.FILES:
                    d = request.FILES['archivo']
                    if d.size > 10485760:
                        return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 10 Mb."})
                    newfiles = request.FILES['archivo']
                    newfilesd = newfiles._name
                    ext = newfilesd[newfilesd.rfind("."):]
                    # extension = newfilesd._name.split('.')
                    # tam = len(extension)
                    # ext = extension[tam - 1]
                    if ext != '.xls' and ext != '.xlsx':
                        return JsonResponse({"result": "bad", "mensaje": u"Error, solo archivos .xls, .xlsx"})
                if f.is_valid():
                    if 'archivo' in request.FILES:
                        newfile = request.FILES['archivo']
                        newfile._name = generar_nombre("archivonotas_", newfile._name)
                    if ArchivoNotas.objects.filter(status=True, descripcion=f.cleaned_data['descripcion']).exists():
                        return JsonResponse({"result": "bad", "mensaje": u"La descripcion ya existe."})
                    archivonotas = ArchivoNotas(descripcion=f.cleaned_data['descripcion'],
                                                observacion=f.cleaned_data['observacion'],
                                                archivo=newfile)
                    archivonotas.save(request)
                    log(u'Adiciono Archivo Notas: %s' % archivonotas, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                     raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                translator = Translator()
                return JsonResponse({"result": "bad", "ex": translator.translate(ex.__str__(),'es').text, "mensaje": u"Error al guardar los datos."})

        elif action == 'edit':
            try:
                archivonotas = ArchivoNotas.objects.get(pk=request.POST['id'])
                f = ArchivoNotasForm(request.POST, request.FILES)
                if 'archivo' in request.FILES:
                    d = request.FILES['archivo']
                    if d.size > 10485760:
                        return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 10 Mb."})
                    newfiles = request.FILES['archivo']
                    newfilesd = newfiles._name
                    ext = newfilesd[newfilesd.rfind("."):]
                    # extension = newfilesd._name.split('.')
                    # tam = len(extension)
                    # ext = extension[tam - 1]
                    if ext != '.xls' and ext != '.xlsx':
                        return JsonResponse({"result": "bad", "mensaje": u"Error, solo archivos .xls, .xlsx"})
                if f.is_valid():
                    if ArchivoNotas.objects.filter(status=True, descripcion=f.cleaned_data['descripcion']).exclude(pk=archivonotas.id).exists():
                        return JsonResponse({"result": "bad", "mensaje": u"La descripcion ya existe."})
                    archivonotas.descripcion = f.cleaned_data['descripcion']
                    archivonotas.observacion = f.cleaned_data['observacion']
                    if 'archivo' in request.FILES:
                        newfile = request.FILES['archivo']
                        newfile._name = generar_nombre("archivonotas_", newfile._name)
                        archivonotas.archivo = newfile
                    archivonotas.save(request)
                    log(u'Modifico Archivo Notas: %s' % archivonotas, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                     raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                translator = Translator()
                return JsonResponse({"result": "bad", "ex": translator.translate(ex.__str__(),'es').text, "mensaje": u"Error al guardar los datos."})

        elif action == 'subirarchivo':
            try:
                archivonotas = ArchivoNotas.objects.get(pk=request.POST['id'])
                f = ArchivoNotasSistemaForm(request.POST, request.FILES)
                if 'archivosistema' in request.FILES:
                    d = request.FILES['archivosistema']
                    if d.size > 10485760:
                        return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 10 Mb."})
                    newfiles = request.FILES['archivosistema']
                    newfilesd = newfiles._name
                    ext = newfilesd[newfilesd.rfind("."):]
                    # extension = newfilesd._name.split('.')
                    # tam = len(extension)
                    # ext = extension[tam - 1]
                    if ext != '.xls' and ext != '.xlsx':
                        return JsonResponse({"result": "bad", "mensaje": u"Error, solo archivos .xls, .xlsx"})
                if f.is_valid():
                    if 'archivosistema' in request.FILES:
                        newfile = request.FILES['archivosistema']
                        newfile._name = generar_nombre("archivonotassistema_", newfile._name)
                        archivonotas.archivosistema = newfile
                    archivonotas.save(request)
                    log(u'Subio Archivo Sistema: %s' % archivonotas, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                     raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                translator = Translator()
                return JsonResponse({"result": "bad", "ex": translator.translate(ex.__str__(),'es').text, "mensaje": u"Error al guardar los datos."})

        elif action == 'delete':
            try:
                archivonotas = ArchivoNotas.objects.get(pk=request.POST['id'])
                log(u'Elimino Archivo Notas: %s' % archivonotas, request, "del")
                archivonotas.delete()
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                translator = Translator()
                return JsonResponse({"result": "bad", "ex": translator.translate(ex.__str__(),'es').text, "mensaje": u"Error al eliminar los datos."})

        elif action == 'procesar':
            try:
                archivonotas = ArchivoNotas.objects.get(pk=request.POST['id'])
                workbook = load_workbook(archivonotas.archivosistema.file.name)
                workbookincidencias = xlwt.Workbook()
                sheetincidencias = workbookincidencias.add_sheet('Incidencias')
                # sheetcorrectos = workbookincidencias.add_sheet('Activosguardados')
                sheet = workbook._sheets[0]
                listaactas = []
                linealectura = 1
                linea=1
                errorfichero = False
                for rowx in sheet.rows:
                    errores = False
                    if linealectura >= 2:
                        id_carrera = str(sheet.cell(row=linealectura, column=1).value)
                        cedula = str(sheet.cell(row=linealectura, column=2).value)
                        id_asignatura = str(sheet.cell(row=linealectura, column=16).value)
                        if not Persona.objects.filter((Q(cedula__icontains=cedula) | Q(pasaporte__icontains=cedula))).exists():
                            sheetincidencias.write(linea, 0, cedula)
                            sheetincidencias.write(linea, 1, 'NO EXISTE LA CEDULA O PASAPORTE')
                            errores = True
                            errorfichero = True
                            linea += 1
                        if not Inscripcion.objects.filter((Q(persona__cedula__icontains=cedula) | Q(persona__pasaporte__icontains=cedula)), carrera_id=int(id_carrera), status=True).exists():
                            sheetincidencias.write(linea, 0, cedula)
                            sheetincidencias.write(linea, 1, u'NO EXISTE LA PERSONA CON ESA CARRERA - %s' % (id_carrera))
                            errores = True
                            errorfichero = True
                            linea += 1
                        if not AsignaturaMalla.objects.filter(malla__carrera_id=int(id_carrera), asignatura_id=int(id_asignatura) , status=True).exists():
                            if not ModuloMalla.objects.filter(malla__carrera_id=int(id_carrera), asignatura_id=int(id_asignatura) , status=True).exists():
                                sheetincidencias.write(linea, 0, cedula)
                                sheetincidencias.write(linea, 1, u'ASIGNATURA NO ESTA EN MALLA O MODULO - %s EN LA CARRERA: %s' % (id_asignatura, id_carrera))
                                errores = True
                                errorfichero = True
                                linea += 1
                    linealectura += 1
                nombre = 'archivonotasincidencias' + str(archivonotas.id) + '.xls'
                nombrearchivo = os.path.join(MEDIA_ROOT, 'archivonotas', nombre)
                workbookincidencias.save(nombrearchivo)
                archivonotas.archivoincidencia.name = "archivonotas/%s" % nombre
                archivonotas.save(request)
                if errorfichero:
                    return JsonResponse({"result": "ok", "obs": True, "archivo": archivonotas.archivoincidencia.url})
                # ingreso
                linea = 1
                contadorguardados = 0
                for rowx in sheet.rows:
                    if linea >= 2:
                        id_carrera = str(sheet.cell(row=linea, column=1).value)
                        cedula = str(sheet.cell(row=linea, column=2).value)
                        id_asignatura = str(sheet.cell(row=linea, column=16).value)
                        fecha = str(sheet.cell(row=linea, column=13).value)
                        fecha_inicio = str(sheet.cell(row=linea, column=14).value)
                        fecha_fin = str(sheet.cell(row=linea, column=15).value)
                        estado_nota = str(sheet.cell(row=linea, column=12).value)[0:1]
                        asistencia = str(sheet.cell(row=linea, column=11).value)
                        nota = str(sheet.cell(row=linea, column=10).value)
                        credito = null_to_decimal(str(sheet.cell(row=linea, column=17).value),2)
                        hora = null_to_decimal(str(sheet.cell(row=linea, column=18).value),2)
                        aprobada = False
                        if estado_nota == 'A':
                            aprobada = True
                        creditos = credito
                        horas = hora
                        modulo = True
                        validapromedio = False
                        asignaturamalla = AsignaturaMalla.objects.filter(malla__carrera_id=int(id_carrera), asignatura_id=int(id_asignatura), status=True)
                        if asignaturamalla:
                            creditos = asignaturamalla[0].creditos
                            horas = asignaturamalla[0].horas
                            modulo = False
                            validapromedio = True
                        inscripcion = Inscripcion.objects.get((Q(persona__cedula__icontains=cedula) | Q(persona__pasaporte__icontains=cedula)),carrera_id=int(id_carrera), status=True)
                        recordaux = RecordAcademico.objects.filter(inscripcion=inscripcion, asignatura_id=int(id_asignatura), status=True)
                        if not recordaux:
                            if modulo:
                                record1 = RecordAcademico(inscripcion=inscripcion,
                                                          asignatura_id=int(id_asignatura),
                                                          nota=int(nota),
                                                          asistencia=int(asistencia),
                                                          fecha=convertirfecha2(fecha),
                                                          fechainicio=convertirfecha2(fecha_inicio),
                                                          fechafin=convertirfecha2(fecha_fin),
                                                          convalidacion=False,
                                                          aprobada=aprobada,
                                                          pendiente=False,
                                                          creditos=creditos,  # preguntar
                                                          horas=horas,  # preguntar
                                                          homologada=False,
                                                          valida=True,
                                                          validapromedio=validapromedio,
                                                          observaciones=archivonotas.observacion)
                            else:
                                record1 = RecordAcademico(inscripcion=inscripcion,
                                                          asignatura_id=int(id_asignatura),
                                                          nota=int(nota),
                                                          asistencia=int(asistencia),
                                                          fecha=convertirfecha2(fecha),
                                                          convalidacion=False,
                                                          aprobada=aprobada,
                                                          pendiente=False,
                                                          creditos=creditos,  # preguntar
                                                          horas=horas,  # preguntar
                                                          homologada=False,
                                                          valida=True,
                                                          validapromedio=validapromedio,
                                                          observaciones=archivonotas.observacion)
                            record1.save()
                            record1.actualizar()
                        else:
                            if recordaux.filter(fecha=convertirfecha2(fecha)).exists():
                                historico = HistoricoRecordAcademico.objects.filter(inscripcion=inscripcion, asignatura_id=int(id_asignatura),fecha=convertirfecha2(fecha))
                                if historico:
                                    historico[0].asignatura_id = int(id_asignatura)
                                    historico[0].asistencia = int(asistencia)
                                    historico[0].nota = int(nota)
                                    historico[0].aprobada = aprobada
                                    historico[0].fecha = convertirfecha2(fecha)
                                    historico[0].creditos = creditos
                                    historico[0].horas = horas
                                    historico[0].observaciones = archivonotas.observacion
                                    historico[0].status = True
                                    historico[0].save(request)
                                    historico[0].actualizar()
                                else:
                                    hrecord1 = HistoricoRecordAcademico(recordacademico=recordaux[0],
                                                                        inscripcion=inscripcion,
                                                                        asignatura_id=int(id_asignatura),
                                                                        nota=int(nota),
                                                                        asistencia=int(asistencia),
                                                                        fecha=convertirfecha2(fecha),
                                                                        convalidacion=False,
                                                                        aprobada=aprobada,
                                                                        pendiente=False,
                                                                        creditos=creditos,  # preguntar
                                                                        horas=horas,  # preguntar
                                                                        homologada=False,
                                                                        valida=True,
                                                                        observaciones=archivonotas.observacion)
                                    hrecord1.save()
                                    hrecord1.actualizar()
                            else:
                                hrecord1 = HistoricoRecordAcademico(recordacademico=recordaux[0],
                                                                    inscripcion=inscripcion,
                                                                    asignatura_id=int(id_asignatura),
                                                                    nota=int(nota),
                                                                    asistencia=int(asistencia),
                                                                    fecha=convertirfecha2(fecha),
                                                                    convalidacion=False,
                                                                    aprobada=aprobada,
                                                                    pendiente=False,
                                                                    creditos=creditos,  # preguntar
                                                                    horas=horas,  # preguntar
                                                                    homologada=False,
                                                                    valida=True,
                                                                    observaciones=archivonotas.observacion)
                                hrecord1.save()
                                hrecord1.actualizar()
                    linea += 1

                archivonotas.procesado = True
                archivonotas.save(request)
                log(u'Proceso archivo de notas: %s' % archivonotas, request, "edit")
                return JsonResponse({"result": "ok", "obs": False, "actas": listaactas})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})


        return JsonResponse({"result": "bad", "mensaje": u"Solicitud Incorrecta."})
    else:
        data['title'] = u'Subir Actas de Notas'
        if 'action' in request.GET:
            action = request.GET['action']

            if action == 'add':
                try:
                    data['title'] = u'Adicionar archivo notas'
                    data['form'] = ArchivoNotasForm()
                    return render(request, "adm_archivonotas/add.html", data)
                except Exception as ex:
                    pass

            elif action == 'edit':
                try:
                    data['title'] = u'Editar archivo notas'
                    archivonotas = ArchivoNotas.objects.get(pk=request.GET['id'])
                    f = ArchivoNotasForm(initial={'descripcion': archivonotas.descripcion,
                                                  'observacion': archivonotas.observacion})
                    data['form'] = f
                    data['archivonotas'] = archivonotas
                    return render(request, "adm_archivonotas/edit.html", data)
                except Exception as ex:
                    pass

            elif action == 'subirarchivo':
                try:
                    data['title'] = u'Subir archivo notas'
                    archivonotas = ArchivoNotas.objects.get(pk=request.GET['id'])
                    f = ArchivoNotasSistemaForm()
                    data['form'] = f
                    data['archivonotas'] = archivonotas
                    return render(request, "adm_archivonotas/subirarchivo.html", data)
                except Exception as ex:
                    pass

            elif action == 'delete':
                try:
                    data['title'] = u'Eliminar archivo notas'
                    data['archivonotas'] = ArchivoNotas.objects.get(pk=request.GET['id'])
                    return render(request, "adm_archivonotas/delete.html", data)
                except Exception as ex:
                    pass

            elif action == 'procesar':
                try:
                    data['title'] = u'Procesar archivo notas'
                    data['archivonotas'] = ArchivoNotas.objects.get(pk=request.GET['id'])
                    return render(request, "adm_archivonotas/procesar.html", data)
                except Exception as ex:
                    pass

            return HttpResponseRedirect(request.path)
        else:
            search = None
            ids = None
            if 's' in request.GET:
                search = request.GET['s']
                ss = search.split(' ')
                if len(ss) == 2:
                    archivonotas = ArchivoNotas.objects.filter(Q(descripcion__icontains=ss[0]) & Q(descripcion__icontains=ss[1]),status=True).order_by('-id')
                if len(ss) == 3:
                    archivonotas = ArchivoNotas.objects.filter(Q(descripcion__icontains=ss[0]) & Q(descripcion__icontains=ss[1]) & Q(descripcion__icontains=ss[2]),status=True).order_by('-id')
                if len(ss) == 4:
                    archivonotas = ArchivoNotas.objects.filter(Q(descripcion__icontains=ss[0]) & Q(descripcion__icontains=ss[1]) & Q(descripcion__icontains=ss[2]) & Q(descripcion__icontains=ss[3]),status=True).order_by('-id')
                else:
                    archivonotas = ArchivoNotas.objects.filter(descripcion__icontains=search, status=True).order_by('-id')
            elif 'id' in request.GET:
                ids = request.GET['id']
                archivonotas = ArchivoNotas.objects.filter(id=ids)
            else:
                archivonotas = ArchivoNotas.objects.filter(status=True).order_by('-id')
            # archivonotas=archivonotas.filter(usuario_creacion=usuario_ingreso)
            paging = MiPaginador(archivonotas, 25)
            p = 1
            try:
                paginasesion = 1
                if 'paginador' in request.session:
                    paginasesion = int(request.session['paginador'])
                if 'page' in request.GET:
                    p = int(request.GET['page'])
                else:
                    p = paginasesion
                try:
                    page = paging.page(p)
                except:
                    p = 1
                page = paging.page(p)
            except:
                page = paging.page(p)
            request.session['paginador'] = p
            data['paging'] = paging
            data['rangospaging'] = paging.rangos_paginado(p)
            data['page'] = page
            data['search'] = search if search else ""
            data['ids'] = ids if ids else ""
            data['archivonotas'] = page.object_list
            data['permiso'] = permiso
            return render(request, "adm_archivonotas/view.html", data)
