# -*- coding: latin-1 -*-
import json
import os
import random
#excel
import sys
import zipfile

from openpyxl import workbook as openxl
from openpyxl.chart import ScatterChart, Reference, Series,PieChart, BarChart
from openpyxl.styles import Font as openxlFont
from openpyxl.styles.alignment import Alignment as alin

from xlwt import *
from datetime import datetime, timedelta

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Q, Count
from django.forms import model_to_dict
from django.http import HttpResponseRedirect, HttpResponse, JsonResponse
from django.shortcuts import render, redirect
from django.template import Context
from django.template.loader import get_template

from decorators import last_access, secure_module, change_periodo_perfil
from settings import SITE_STORAGE
from sga.commonviews import adduserdata
from sga.forms import AperturaPeriodoCambioCarreraForm, DocumentoRequeridoCambioCarreraForm, \
    SubirEvidenciaCambioCarreraForm
from sga.funciones import log, MiPaginador, generar_nombre, notificacion, \
    remover_caracteres_especiales_unicode, validar_archivo
from sga.models import PracticasPreprofesionalesInscripcion, CUENTAS_CORREOS, Carrera, \
    AperturaPracticaPreProfesional, \
    CoordinadorCarrera, ESTADOS_PASOS_SOLICITUD, ResponsableCoordinacion, AperturaPeriodoCambioCarrera, \
    RequisitosCambioCarrera, \
    ESTADO_SOLICITUD_CAMBIO_CARRERA, SolicitudCambioCarrera, \
    DocumentosSolicitudCambioCarrera, CarrerasCambioCarrera, SeguimientoRevisionDocumentoCC, ConvalidaCambioCarrera, \
    Persona, ReponsableCambioCarrera, ROLES_CAMBIO_CARRERA, Administrativo, DocumentosValidados, Notificacion, \
    HistorialDocumentosSolicitudCC, Malla, Materia
from sga.tasks import send_html_mail
from sga.templatetags.sga_extras import encrypt



@login_required(redirect_field_name='ret', login_url='/loginsga')
@change_periodo_perfil
@secure_module
@last_access
@transaction.atomic()
def view(request):
    global ex
    persona = request.session['persona']
    coordinacion = []
    data = {}
    perfilprincipal = request.session['perfilprincipal']
    periodo = request.session['periodo']
    if perfilprincipal.es_estudiante():
        return HttpResponseRedirect("/?info=Solo los perfiles administrativo y docente pueden ingresar al modulo.")
    if persona.es_profesor():
        coordinacion = persona.profesor().coordinacion
    if persona.id in [17579, 818, 5194, 23532, 169, 12130, 16630, 1652, 21604, 30751, 30802, 27946, 16781]:
        coordinacion = []

    data['miscarreras'] = miscarreras = persona.mis_carreras_tercer_nivel()
    tiene_carreras_director = True if miscarreras else False
    data['querydecano'] = querydecano = ResponsableCoordinacion.objects.filter(periodo=periodo, status=True, coordinacion_id__lte=5,
                                                                               persona=persona, tipo=1)
    es_director_carr = tiene_carreras_director if not querydecano.exists() else False
    data['es_director_carr'] = es_director_carr
    data['es_decano'] = es_decano = querydecano.exists()

    responsables_admision=ReponsableCambioCarrera.objects.filter(status=True)
    personas_admision = responsables_admision.filter(rol=0)
    data['es_asist_bienestar']=es_asist_bienestar=responsables_admision.filter(persona=persona, rol=1).exists()
    data['es_director_adm']=es_director_adm=responsables_admision.filter(persona=persona, rol=2).exists()

    if request.method == 'POST':
        action = request.POST['action']

        if action == 'loadDataSolicitudes':
            try:
                if not 'id' in request.POST:
                    raise NameError("Parametro de periodo no encontrado")
                if not AperturaPeriodoCambioCarrera.objects.filter(pk=int(request.POST['id'])).exists():
                    raise NameError("Periodo no encontrado")
                aPeriodo = AperturaPeriodoCambioCarrera.objects.get(pk=int(request.POST['id']))
                if not SolicitudCambioCarrera.objects.filter(periodocambiocarrera=aPeriodo, status=True).exists():
                    raise NameError("No existen solicitudes para graficar")

                solicitudes= SolicitudCambioCarrera.objects.filter(periodocambiocarrera=aPeriodo, status=True)
                aaData = {}
                tipografico=1
                if 'tipografico' in request.POST:
                    tipografico=int(request.POST['tipografico'])
                aaData['tipografico'] =tipografico
                aaData['total_solicitudes'] = total_solicitudes = len(solicitudes)
                solivariables = []
                if tipografico == 1:
                    aaData['total_cambioies'] = total_cambioies = len(solicitudes.filter(inscripcion__isnull=True))
                    aaData['total_cambiocarrera'] = total_cambiocarrera = len(solicitudes.filter(inscripcion__isnull=False))
                    solivariables.append({"nombre": "SOLICITUDES DE CAMBIO DE CARRERA",
                                           "alias": "CAMBIO CARRERA",
                                           "total": total_cambiocarrera})
                    solivariables.append({"nombre": "SOLICITUDES DE CAMBIO DE IES",
                                           "alias": "CAMBIO IES",
                                           "total": total_cambioies})
                if tipografico == 2:
                    estados=ESTADO_SOLICITUD_CAMBIO_CARRERA
                    for estado in estados:
                        aaData[estado[1].replace(" ", "")]=len(solicitudes.filter(estados=estado[0]))
                        aaData[estado[1].replace(" POR ", "")+'C'] = len(solicitudes.filter(estados=estado[0], inscripcion__isnull=False))
                        aaData[estado[1].replace(" POR ", "")+'I'] = len(solicitudes.filter(estados=estado[0], inscripcion__isnull=True))
                        solivariables.append({"nombre":estado[1],
                                              "alias":estado[1].replace(" POR ", " "),
                                              "cambiocarrera": len(solicitudes.filter(estados=estado[0], inscripcion__isnull=False)),
                                              "cambioies": len(solicitudes.filter(estados=estado[0], inscripcion__isnull = True)),
                                              "total":len(solicitudes.filter(estados=estado[0]))})
                return JsonResponse({"result": "ok", "data": aaData, "coordinaciones": solivariables})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al cargar los datos. %s" % ex.__str__(), "data": [], "iTotalRecords": 0, "iTotalDisplayRecords": 0})

        if action == 'adddocumento':
            try:
                with transaction.atomic():
                    form = DocumentoRequeridoCambioCarreraForm(request.POST, request.FILES)
                    if form.is_valid():
                        filtro = RequisitosCambioCarrera(nombre=form.cleaned_data['nombre'].upper(),
                                                         leyenda=form.cleaned_data['leyenda'],
                                                         multiple=form.cleaned_data['multiple'],
                                                         externo=form.cleaned_data['externo'],
                                                         opcional=form.cleaned_data['opcional'],
                                                         essilabo=form.cleaned_data['essilabo'],
                                                         )
                        filtro.save(request)
                        if 'archivo' in request.FILES:
                            newfile = request.FILES['archivo']
                            newfile._name = generar_nombre(filtro.nombre_input(), newfile._name)
                            filtro.archivo = newfile
                            filtro.save(request)
                        log(u'Adiciono Documento de Cambio de Carrera: %s' % filtro, request, "add")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Complete los datos requeridos."}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        if action == 'editdocumento':
            try:
                with transaction.atomic():
                    filtro = RequisitosCambioCarrera.objects.get(pk=request.POST['id'])
                    f = DocumentoRequeridoCambioCarreraForm(request.POST, request.FILES)
                    if f.is_valid():
                        filtro.nombre = f.cleaned_data['nombre'].upper()
                        filtro.leyenda = f.cleaned_data['leyenda']
                        filtro.documentorequerido = f.cleaned_data['documentorequerido']
                        filtro.multiple = f.cleaned_data['multiple']
                        filtro.externo = f.cleaned_data['externo']
                        filtro.opcional=f.cleaned_data['opcional']
                        filtro.essilabo = f.cleaned_data['essilabo']
                        if 'archivo' in request.FILES:
                            newfile = request.FILES['archivo']
                            newfile._name = generar_nombre(filtro.nombre_input(), newfile._name)
                            filtro.archivo = newfile
                        filtro.save(request)
                        log(u'Modificó Documento de Cambio de Carrera: %s' % filtro, request, "edit")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                             "mensaje": "Complete los datos requeridos."}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        if action == 'deletedocumento':
            try:
                with transaction.atomic():
                    instancia = RequisitosCambioCarrera.objects.get(pk=int(request.POST['id']))
                    instancia.status = False
                    instancia.save(request)
                    log(u'Elimino Documento de Cambio de Carrera: %s' % instancia, request, "delete")
                    res_json = {"error": False}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        elif action == 'addaperturaperiodo':
            try:
                f = AperturaPeriodoCambioCarreraForm(request.POST)
                if f.is_valid():
                    aperturas=AperturaPeriodoCambioCarrera.objects.filter(status=True)
                    if f.cleaned_data['fechaapertura'] < datetime.now().date():
                        return JsonResponse({"result": "bad",
                                             "mensaje": "La fecha de inicio tiene que ser mayor o igual a la fecha actual."})
                    if aperturas:
                        for apertura in aperturas:
                            if apertura.fechacierre >= f.cleaned_data['fechaapertura'] >= apertura.fechaapertura:
                                return JsonResponse({"result": "bad",
                                                     "mensaje": "Ya existe un periodo aperturado con el rango de fechas seleccionado."})
                    if not f.cleaned_data['fechaapertura'] <= f.cleaned_data['fechacierre']:
                        return JsonResponse({"result": "bad",
                                             "mensaje": "La fecha de apertura no puede ser mayor a la fecha de cierre."})
                    if not f.cleaned_data['fechainiciorequisitosadmision'] <= f.cleaned_data['fechacierrerequisitosadmision']:
                        return JsonResponse({"result": "bad", "mensaje": "La fecha de inicio de revision de admision no puede ser mayor a la fecha de cierre de revision de admision."})
                    elif f.cleaned_data['fechainiciorecepciondocumentos'] < f.cleaned_data['fechaapertura'] or \
                            f.cleaned_data['fechacierrerecepciondocumentos'] > f.cleaned_data['fechacierre']:
                        return JsonResponse({"result": "bad",
                                             "mensaje": "La fecha de recepción de documentos fuera de rango."})
                    elif f.cleaned_data['fechainiciorequisitosadmision'] < f.cleaned_data['fechaapertura'] or \
                            f.cleaned_data['fechacierrerequisitosadmision'] > f.cleaned_data['fechacierre']:
                        return JsonResponse({"result": "bad",
                                             "mensaje": "La fecha de revison de requisitos de admision  fuera de rango."})
                    if not f.cleaned_data['fechainicioremitirdecano'] <= f.cleaned_data['fechacierreremitirdecano']:
                        return JsonResponse({"result": "bad",
                                             "mensaje": "La fecha de inicio de revision del decano no puede ser mayor a la fecha de cierre de revision del decano."})
                    elif f.cleaned_data['fechainicioremitirdecano'] < f.cleaned_data['fechaapertura'] or f.cleaned_data[
                        'fechacierreremitirdecano'] > f.cleaned_data['fechacierre']:
                        return JsonResponse(
                            {"result": "bad", "mensaje": "La fecha de revison del decano fuera de rango."})
                    if not f.cleaned_data['fechainiciovaldirector'] <= f.cleaned_data['fechacierrevaldirector']:
                        return JsonResponse({"result": "bad",
                                             "mensaje": "La fecha de inicio de revision del director no puede ser mayor a la fecha de cierre de revision del director."})
                    elif f.cleaned_data['fechainiciovaldirector'] < f.cleaned_data['fechaapertura'] or f.cleaned_data[
                        'fechacierrevaldirector'] > f.cleaned_data['fechacierre']:
                        return JsonResponse(
                            {"result": "bad", "mensaje": "La fecha de revison del director fuera de rango."})
                    if not f.cleaned_data['fechainicioremitiraprobados'] <= f.cleaned_data['fechacierreremitiraprobados']:
                        raise NameError("La fecha de inicio de remición de aprobados no puede ser mayor a la fecha de cierre de remición de aprobados.")
                    elif f.cleaned_data['fechainicioremitiraprobados'] < f.cleaned_data['fechaapertura'] or f.cleaned_data[
                        'fechacierreremitiraprobados'] > f.cleaned_data['fechacierre']:
                       raise NameError("La fecha de remitir aprobados fuera de rango.")
                    if AperturaPeriodoCambioCarrera.objects.filter(fechaapertura=f.cleaned_data['fechaapertura'],
                                                                   fechacierre=f.cleaned_data['fechacierre'],
                                                                   status=True).exists():
                        return JsonResponse(
                            {"result": "bad", "mensaje": "Ya existe un registro con las misma fechas indicadas"})
                    if not f.cleaned_data['mensaje']:
                        return JsonResponse({"result": "bad", "mensaje": "Ingrese un mensaje"})
                    apertura = AperturaPeriodoCambioCarrera(mensaje=f.cleaned_data['mensaje'],
                                                            fechaapertura=f.cleaned_data['fechaapertura'],
                                                            fechacierre=f.cleaned_data['fechacierre'],
                                                            # periodo=f.cleaned_data['periodo'],
                                                            motivo=f.cleaned_data['motivo'],
                                                            fechainiciorecepciondocumentos = f.cleaned_data['fechainiciorecepciondocumentos'],
                                                            fechacierrerecepciondocumentos = f.cleaned_data['fechacierrerecepciondocumentos'],
                                                            fechainiciorequisitosadmision=f.cleaned_data[
                                                                'fechainiciorequisitosadmision'],
                                                            fechacierrerequisitosadmision=f.cleaned_data[
                                                                'fechacierrerequisitosadmision'],
                                                            fechainicioremitirdecano=f.cleaned_data[
                                                                'fechainicioremitirdecano'],
                                                            fechacierreremitirdecano=f.cleaned_data[
                                                                'fechacierreremitirdecano'],
                                                            fechainiciovaldirector=f.cleaned_data[
                                                                'fechainiciovaldirector'],
                                                            fechacierrevaldirector=f.cleaned_data[
                                                                'fechacierrevaldirector'],
                                                            fechainicioremitiraprobados=f.cleaned_data[
                                                                'fechainicioremitiraprobados'],
                                                            fechacierreremitiraprobados=f.cleaned_data[
                                                                'fechacierreremitiraprobados'],
                                                            )
                    apertura.save()
                    for data in f.cleaned_data['coordinacion']:
                        apertura.coordinacion.add(data)
                    for data in f.cleaned_data['requisitos']:
                        apertura.requisitos.add(data)
                    # apertura.publico = f.cleaned_data['publico']
                    apertura.save(request)
                    log(u'Adiciono apertura de practica pre profesional: %s - %s - %s' % (
                    apertura.fechaapertura, apertura.fechacierre, apertura.motivo), request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": str(ex)})

        elif action == 'editaperturaperiodo':
            try:
                f = AperturaPeriodoCambioCarreraForm(request.POST)
                if f.is_valid():
                    apertura = AperturaPeriodoCambioCarrera.objects.get(pk=int(request.POST['id']))
                    if not apertura.fechaapertura <= apertura.fechacierre:
                        return JsonResponse({"result": "bad",
                                             "mensaje": "La fecha de apertura no puede ser mayor a la fecha de cierre."})
                    if f.cleaned_data['fechainiciorecepciondocumentos'] < f.cleaned_data['fechaapertura'] or \
                            f.cleaned_data['fechacierrerecepciondocumentos'] > f.cleaned_data['fechacierre']:
                        return JsonResponse({"result": "bad", "mensaje": "La fecha de recepción de documentos fuera de rango."})
                    if not f.cleaned_data['fechainiciorequisitosadmision'] <= f.cleaned_data[
                        'fechacierrerequisitosadmision']:
                        return JsonResponse({"result": "bad",
                                             "mensaje": "La fecha de inicio de revision de admision no puede ser mayor a la fecha de cierre de revision de admision."})
                    elif f.cleaned_data['fechainiciorequisitosadmision'] < f.cleaned_data['fechaapertura'] or \
                            f.cleaned_data['fechacierrerequisitosadmision'] > f.cleaned_data['fechacierre']:
                        return JsonResponse({"result": "bad",
                                             "mensaje": "La fecha de revison de requisitos de admision  fuera de rango."})
                    if not f.cleaned_data['fechainicioremitirdecano'] <= f.cleaned_data['fechacierreremitirdecano']:
                        return JsonResponse({"result": "bad",
                                             "mensaje": "La fecha de inicio de revision del decano no puede ser mayor a la fecha de cierre de revision del decano."})
                    elif f.cleaned_data['fechainicioremitirdecano'] < f.cleaned_data['fechaapertura'] or f.cleaned_data[
                        'fechacierreremitirdecano'] > f.cleaned_data['fechacierre']:
                        return JsonResponse(
                            {"result": "bad", "mensaje": "La fecha de revison del decano fuera de rango."})
                    if not f.cleaned_data['fechainiciovaldirector'] <= f.cleaned_data['fechacierrevaldirector']:
                        return JsonResponse({"result": "bad",
                                             "mensaje": "La fecha de inicio de revision del director no puede ser mayor a la fecha de cierre de revision del director."})
                    elif f.cleaned_data['fechainiciovaldirector'] < f.cleaned_data['fechaapertura'] or f.cleaned_data[
                        'fechacierrevaldirector'] > f.cleaned_data['fechacierre']:
                        return JsonResponse(
                            {"result": "bad", "mensaje": "La fecha de revison del director fuera de rango."})
                    if not f.cleaned_data['fechainicioremitiraprobados'] <= f.cleaned_data['fechacierreremitiraprobados']:
                        raise NameError("La fecha de inicio de remición de aprobados no puede ser mayor a la fecha de cierre de remición de aprobados.")
                    elif f.cleaned_data['fechainicioremitiraprobados'] < f.cleaned_data['fechaapertura'] or f.cleaned_data[
                        'fechacierreremitiraprobados'] > f.cleaned_data['fechacierre']:
                       raise NameError("La fecha de remitir aprobados fuera de rango.")
                    if AperturaPeriodoCambioCarrera.objects.filter(fechaapertura=f.cleaned_data['fechaapertura'],
                                                                   fechacierre=f.cleaned_data['fechacierre']).exclude(
                            pk=int(request.POST['id'])).exists():
                        return JsonResponse(
                            {"result": "bad", "mensaje": "Ya existe un registro con las misma fechas indicadas"})
                    if not f.cleaned_data['mensaje']:
                        return JsonResponse({"result": "bad", "mensaje": "Ingrese un mensaje"})
                    apertura.mensaje = f.cleaned_data['mensaje']
                    apertura.fechaapertura = f.cleaned_data['fechaapertura']
                    apertura.fechacierre = f.cleaned_data['fechacierre']
                    # apertura.periodo = f.cleaned_data['periodo']
                    apertura.motivo = f.cleaned_data['motivo']
                    # apertura.publico = f.cleaned_data['publico']
                    apertura.fechainiciorecepciondocumentos = f.cleaned_data['fechainiciorecepciondocumentos']
                    apertura.fechacierrerecepciondocumentos = f.cleaned_data['fechacierrerecepciondocumentos']
                    apertura.fechainiciorequisitosadmision = f.cleaned_data['fechainiciorequisitosadmision']
                    apertura.fechacierrerequisitosadmision = f.cleaned_data['fechacierrerequisitosadmision']
                    apertura.fechainicioremitirdecano = f.cleaned_data['fechainicioremitirdecano']
                    apertura.fechacierreremitirdecano = f.cleaned_data['fechacierreremitirdecano']
                    apertura.fechainiciovaldirector = f.cleaned_data['fechainiciovaldirector']
                    apertura.fechacierrevaldirector = f.cleaned_data['fechacierrevaldirector']
                    apertura.fechainicioremitiraprobados = f.cleaned_data['fechainicioremitiraprobados']
                    apertura.fechacierreremitiraprobados = f.cleaned_data['fechacierreremitiraprobados']
                    apertura.save(request)
                    apertura.coordinacion.clear()
                    apertura.requisitos.clear()
                    for cord in f.cleaned_data['coordinacion']:
                        apertura.coordinacion.add(cord)
                    for req in f.cleaned_data['requisitos']:
                        apertura.requisitos.add(req)
                    log(u'Edito apertura de periodo de cambio de carrera: %s - %s - %s' % (
                    apertura.fechaapertura, apertura.fechacierre, apertura.motivo), request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "mensaje": "Error en el formulario"})
                    # raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": str(ex)})

        elif action == 'delaperturasolicitud':
            try:
                apertura = AperturaPeriodoCambioCarrera.objects.get(pk=int(request.POST['id']))
                if not apertura.puede_eliminar():
                    return JsonResponse({"result": "bad", "mensaje": "No puede eliminar registro de periodo."})
                for carrera in apertura.carrerascambiocarrera_set.filter(status=True):
                    if carrera.con_documentos():
                        transaction.set_rollback(True)
                        return JsonResponse({"result": "bad", "mensaje": "No puede eliminar registro de periodo."})
                    else:
                        carrera.status = False
                        carrera.save(request)
                apertura.status = False
                apertura.save(request)
                log(u'Elimino periodo cambio carrera: %s - %s - %s' % (
                apertura.fechaapertura, apertura.fechacierre, apertura.motivo), request, "del")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": "Error al eliminar los datos."})

        elif action == 'validardocumentoscambiocarrera':
            try:
                postar = DocumentosSolicitudCambioCarrera.objects.get(id=int(request.POST['id']))
                postar.estados = request.POST['est']
                postar.observacion = request.POST['obs']
                if int(request.POST['est']) == 3:
                    postar.doccorregido = False
                postar.save(request)
                soli = SolicitudCambioCarrera.objects.get(pk=postar.solicitud.pk)
                soli.fecha_revision_admision = datetime.now()
                soli.save(request)
                if int(request.POST['est']) == 3:
                    historial=HistorialDocumentosSolicitudCC(documento=postar,
                                                             fecha_revision=datetime.now(),
                                                             # tiempo=request.POST['time'],
                                                             estados=postar.estados,
                                                             observacion=postar.observacion)
                    historial.save(request)
                asunto = u"VALIDACIÓN DE DOCUMENTOS DE CAMBIO DE CARRERA"
                if int(postar.estados) != 1:
                    soli = SolicitudCambioCarrera.objects.get(pk=postar.solicitud.pk)
                    soli.revision_admision = postar.estados
                    if int(postar.estados) == 4:
                        soli.revision_admision = 3

                    soli.fecha_revision_admision = datetime.now()
                    soli.persona_admision = persona
                    soli.save(request)
                    tipocambio=True
                    if postar.solicitud.inscripcion:
                        para = postar.solicitud.inscripcion.persona
                        notificacion(asunto, postar.observacion, para, None,
                                     '/alu_solicitudcambiocarrera?action=verproceso&id={}'.format(
                                         encrypt(postar.solicitud.pk)), postar.pk, 1, 'sga',
                                     DocumentosSolicitudCambioCarrera, request)
                    else:
                        tipocambio=False
                        titulo = "Solicitud de cambio de IES pendiente de revisión"
                        lista_email = soli.persona.lista_emails(),
                        # lista_email = ['jguachuns@unemi.edu.ec', ]
                        datos_email = {'sistema': request.session['nombresistema'],
                                       'fecha': datetime.now().date(),
                                       'hora': datetime.now().time(),
                                       'tipocambio': tipocambio,
                                       'estado': soli.get_estados_display,
                                       'persona':postar.solicitud.persona,
                                       'asunto':asunto,
                                       'mensaje': 'Te informamos que tienes documentos por corregir en tu solicitud de cambio de IES'}
                        template = "emails/notificacion_cambio_ies.html"
                        send_html_mail(titulo, template, datos_email, lista_email, [], [], cuenta=CUENTAS_CORREOS[0][1])

                seguimiento = SeguimientoRevisionDocumentoCC(revisor=persona, fecha=datetime.now(), documento=postar,
                                                             solictud=postar.solicitud, estados=postar.estados,
                                                             rutaarchivo=postar.archivo.url,
                                                             observacion=postar.observacion)
                seguimiento.save(request)
                log(u'Valido documento cambio carrera estudiante: %s %s' % (postar, postar.solicitud.inscripcion), request, "edit")

                # SE COMENTA PARA SALTAR EL PASO DE BIENESTAR 26-01-2023
                # if postar.solicitud.documentos_aprobados() == 1 and postar.solicitud.revision_bienestar == 0:
                    # verificador = SolicitudCambioCarrera.objects.filter(status=True, periodocambiocarrera=soli.periodocambiocarrera).values('id')
                    # cb = 0
                    # cr = 0
                    # cont = 0
                    # persona_bienestar_id = 0
                    # responsables = ReponsableCambioCarrera.objects.filter(status=True, rol=1, estado=True).values_list('persona_id', flat=True)
                    # resinactivos = ReponsableCambioCarrera.objects.filter(status=True, rol=1, estado=False).values_list('persona_id', flat=True)
                    # if not responsables:
                    #     transaction.set_rollback(True)
                    #     return JsonResponse({"result": "bad", "mensaje": u"No existen Funcionarios de Bienestar configurados para asignar solicitud, configure los funcionarios ."})
                    # # lista = []
                    # for p in responsables:
                    #     cont += 1
                    #     if not verificador.filter(persona_bienestar=p).exists():
                    #         persona_bienestar_id = p
                    #         break
                    # if persona_bienestar_id == 0 and verificador.exists():
                    #     persona_bienestar_id = SolicitudCambioCarrera.objects.filter(status=True, periodocambiocarrera=postar.solicitud.periodocambiocarrera).exclude(persona_bienestar_id__in=resinactivos).values(
                    #         'persona_bienestar_id').annotate(total=Count('id')).order_by('total')[0]['persona_bienestar_id']
                    #     # elif len(verificador) >= len(responsables):
                    #     #     ct = len(verificador.filter(persona_admision=p))
                    #     #     if ct > cr:
                    #     #         th = ct
                    #     #     elif ct < cr or cb == 0:
                    #     #         cb = ct
                    #     #         per = p
                    #     #     elif cr == ct:
                    #     #         if len(lista) != 0:
                    #     #             lista = lista+[p]
                    #     #         else:
                    #     #             lista = lista+[p2, p]
                    #     #     cr = ct
                    #     #     if len(responsables):
                    #     #         persona_bienestar_id = responsables[random.randint(0, 3)]
                    #     #         break
                    #     #     elif cb < th and cont == len(responsables):
                    #     #         persona_bienestar_id = per[0]
                    #     #         break
                    #     #     p2 = p
                    #
                doc=DocumentosSolicitudCambioCarrera.objects.filter(solicitud=soli, status=True)
                if len(doc)==len(doc.filter(estados=1)):
                    soli.revision_admision=1
                    soli.revision_bienestar=1
                    # soli.persona_bienestar_id = int(persona_bienestar_id)
                    soli.save()
                    # notificacion('SOLICITUD DE CAMBIO DE CARRERA PENDIENTE DE REVISIÓN', 'Se le ha asignado una solicitud de cambio de carrera para revisión',
                    #             soli.persona_bienestar, None,
                    #              '/alu_cambiocarrera?action=solicitantes&id={}&search={}'.format(
                    #                  soli.periodocambiocarrera.pk, soli.inscripcion.persona.cedula if soli.inscripcion else soli.persona.cedula ), soli.pk, 1,
                    #              'sga', SolicitudCambioCarrera, request)
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'verificacionrequisitoscambiocarrera':
            try:
                with transaction.atomic():
                    filtro = SolicitudCambioCarrera.objects.get(pk=int(request.POST['id']))
                    filtro.aprobacion_admision = request.POST['estado']
                    filtro.observacion_admision = request.POST['observacion'].upper()
                    filtro.fecha_aprobacion_admision = datetime.now()
                    filtro.persona_director_admision = persona
                    if request.FILES:
                        newfile = request.FILES['archivo']
                        extension = newfile._name.split('.')
                        tam = len(extension)
                        exte = extension[tam - 1]
                        if newfile.size > 4194304:
                            transaction.set_rollback(True)
                            return JsonResponse(
                                {"result": True, "mensaje": u"Error, el tamaño del archivo es mayor a 4 Mb."})
                        if not exte in ['pdf']:
                            transaction.set_rollback(True)
                            return JsonResponse({"result": True, "mensaje": u"Error, solo archivos .pdf,.jpg, .jpeg"})
                        if filtro.inscripcion:
                            apellido1=filtro.inscripcion.persona.apellido1
                            para = filtro.inscripcion.persona
                        else:
                            apellido1=filtro.persona.apellido1
                            para = filtro.persona
                        nombre_persona = remover_caracteres_especiales_unicode(apellido1).lower().replace(' ', '_')
                        newfile._name = generar_nombre(
                            "{}__{}".format(nombre_persona, 'INFORME_ADMISION_CAMBIO_CARRERA'), newfile._name)
                        filtro.archivoinformeadmsion = newfile
                    filtro.save(request)
                    if filtro.aprobacion_admision == '2':
                        filtro.estados = 6
                        messages.success(request, 'Validación de requisitos Rechazada.')
                    else:
                        filtro.estados = 3
                        messages.success(request, 'Validación de requisitos Aprobada.')
                    asunto = u"VALIDACIÓN DE DOCUMENTOS DE CAMBIO DE CARRERA {}".format(filtro.get_estados_display())
                    notificacion(asunto, filtro.observacion_admision, para, None,
                                 '/alu_solicitudcambiocarrera?action=verproceso&id={}'.format(encrypt(filtro.pk)),
                                 filtro.pk, 1, 'sga', SolicitudCambioCarrera, request)

                    # # Notificacion para el decano de la facultad
                    # subject = 'SOLICITUD DE CAMBIO DE CARRERA PENDIENTE DE REVISIÓN'
                    # asuntodecano = 'Tiene una solicitud de cambio de carrera de la estudiante {} pendiente de revision'.format(
                    #     para.__str__())
                    # decano_facultad = filtro.get_decano()
                    # notificacion(subject, asuntodecano, decano_facultad.persona, None,
                    #              '/alu_cambiocarrera?action=solicitantes&id={}&search={}'.format(
                    #                  filtro.periodocambiocarrera.pk, para.cedula), filtro.pk, 1,
                    #              'sga', SolicitudCambioCarrera, request)
                    filtro.save(request)
                    log(u'Departamento de Admision realizo la verificación de requisitos: {} {}'.format(
                        filtro.inscripcion, filtro.get_revision_admision_display()), request, "add")
                    return JsonResponse({"result": False}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'subirdocbienestar':
            try:
                with transaction.atomic():
                    filtro = SolicitudCambioCarrera.objects.get(pk=int(request.POST['id']))
                    filtro.revision_bienestar = request.POST['estado']
                    filtro.fecha_revision_bienestar = datetime.now()
                    filtro.persona_bienestar = persona
                    filtro.observacion_bienestar=request.POST['observacion']
                    if request.FILES:
                        newfile = request.FILES['archivo']
                        extension = newfile._name.split('.')
                        tam = len(extension)
                        exte = extension[tam - 1]
                        if newfile.size > 4194304:
                            transaction.set_rollback(True)
                            return JsonResponse(
                                {"result": True, "mensaje": u"Error, el tamaño del archivo es mayor a 4 Mb."})
                        if not exte in ['pdf']:
                            transaction.set_rollback(True)
                            return JsonResponse({"result": True, "mensaje": u"Error, solo archivos .pdf,.jpg, .jpeg"})

                        if filtro.inscripcion:
                            apellido1=filtro.inscripcion.persona.apellido1
                        else:
                            apellido1=filtro.persona.apellido1
                        nombre_persona = remover_caracteres_especiales_unicode(apellido1).lower().replace(' ', '_')
                        newfile._name = generar_nombre(
                            "{}__{}".format(nombre_persona, 'INFORME_BIENESTAR_CAMBIO_CARRERA'), newfile._name)
                        filtro.archivobienestar = newfile
                    filtro.save(request)
                    log(u'Departamento de Bienestar realizo la verificación de requisitos: {} {}'.format(
                        filtro.inscripcion, filtro.get_revision_bienestar_display()), request, "add")
                    return JsonResponse({"result": False}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'validardecano':
            try:
                with transaction.atomic():
                    filtro = SolicitudCambioCarrera.objects.get(pk=int(request.POST['id']))
                    filtro.revision_decano = request.POST['estado']
                    filtro.observacion_decano = request.POST['observacion'].upper()
                    filtro.fecha_revision_decano = datetime.now()
                    filtro.persona_decano = persona
                    filtro.save(request)
                    if filtro.inscripcion:
                        estudiante = filtro.inscripcion.persona
                    else:
                        estudiante = filtro.persona
                    if filtro.revision_decano == '2':
                        filtro.estados = 8
                        messages.success(request, 'Validación de Decano de la Facultad Rechazada.')
                    else:
                        filtro.estados = 7
                        messages.success(request, 'Validación de Decano de la Facultad Aprobada.')
                        # Notificacion para el director de la carrera
                        subjectdirector = 'SOLICITUDES DE CAMBIO DE CARRERA/IES-MODALIDAD PENDIENTE DE REVISIÓN'
                        asuntodirector = 'Estimado(a) Director(a) Saludos cordiales, tiene una solicitud de cambio de carrera del estudiante {} para su revisión'.format(
                            estudiante.__str__())
                        director = filtro.get_director()
                        notificacion(subjectdirector, asuntodirector, director.persona, None,
                                     '/alu_cambiocarrera?action=solicitantes&id={}&search={}'.format(
                                         filtro.periodocambiocarrera.pk, estudiante.cedula), filtro.pk,
                                     1, 'sga', SolicitudCambioCarrera, request)
                    filtro.save(request)
                    asunto = u"VALIDACIÓN DE CAMBIO DE CARRERA DEL DECANO {}".format(filtro.get_estados_display())
                    para = estudiante
                    if filtro.inscripcion:
                        notificacion(asunto, filtro.observacion_decano, para, None,
                                     '/alu_solicitudcambiocarrera?action=verproceso&id={}'.format(encrypt(filtro.pk)),
                                     filtro.pk, 1, 'sga', SolicitudCambioCarrera, request)
                        filtro.save(request)
                    log(u'Decano validó proceso de Cambio de Carrera: {} {}'.format(filtro.inscripcion if filtro.inscripcion else estudiante,
                                                                                    filtro.get_revision_director_display()),
                        request, "add")
                    return JsonResponse({"result": False}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'resoluciondirector':
            try:
                with transaction.atomic():
                    filtro = SolicitudCambioCarrera.objects.get(pk=int(request.POST['id']))
                    filtro.revision_director = request.POST['estado']
                    if filtro.inscripcion:
                        estudiante = filtro.inscripcion.persona
                    else:
                        estudiante = filtro.persona
                    if request.FILES:
                        newfile = request.FILES['archivo']
                        extension = newfile._name.split('.')
                        tam = len(extension)
                        exte = extension[tam - 1]
                        if newfile.size > 4194304:
                            transaction.set_rollback(True)
                            return JsonResponse(
                                {"result": True, "mensaje": u"Error, el tamaño del archivo es mayor a 4 Mb."})
                        if not exte in ['pdf']:
                            transaction.set_rollback(True)
                            return JsonResponse({"result": True, "mensaje": u"Error, solo archivos .pdf,.jpg, .jpeg"})
                        nombre_persona = remover_caracteres_especiales_unicode(
                            estudiante.apellido1).lower().replace(' ', '_')
                        newfile._name = generar_nombre("{}__{}".format(nombre_persona, 'RESOLUSION_CAMBIO_CARRERA'),
                                                       newfile._name)
                        filtro.archivoresoluciondirector = newfile
                    filtro.observacion_director = request.POST['observacion'].upper()
                    filtro.fecha_revision_director = datetime.now()
                    filtro.persona_director = persona
                    filtro.save(request)
                    if filtro.revision_director == '2':
                        filtro.estados = 7
                        messages.success(request, 'Validación de Director de Carrera Rechazada.')
                    else:
                        filtro.estados = 4
                        messages.success(request, 'Validación de Director de Carrera Aprobada.')
                    filtro.save(request)
                    asunto = u"VALIDACIÓN DE CAMBIO DE CARRERA DEL DIRECTOR {}".format(filtro.get_estados_display())
                    para = estudiante
                    if filtro.inscripcion:
                        notificacion(asunto, filtro.observacion_director, para, None,
                                     '/alu_solicitudcambiocarrera?action=verproceso&id={}'.format(encrypt(filtro.pk)),
                                     filtro.pk, 1, 'sga', SolicitudCambioCarrera, request)
                    filtro.estados = 1
                    filtro.save(request)
                    log(u'Finalizo Validación de Cambio de Carrera: {} {}'.format(filtro.inscripcion if filtro.inscripcion else estudiante,
                                                                                  filtro.get_revision_director_display()),
                        request, "add")
                    return JsonResponse({"result": False}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'finalizarproceso':
            try:
                data['filtro'] = solicitud = SolicitudCambioCarrera.objects.get(pk=int(request.POST['id']))
                solicitud.estados = 2
                solicitud.save(request)
                return JsonResponse({'resp': True, 'mensaje': 'Proceso Finalizado'})
            except Exception as ex:
                return JsonResponse({'resp': False, 'mensaje': str(ex)})

        elif action == 'notidirector':
            try:
                soli = SolicitudCambioCarrera.objects.get(pk=request.POST['id'])
                soli.fecha_notificacion_director = datetime.now()
                soli.save(request)
                if soli.inscripcion:
                    estudiante = soli.inscripcion.persona
                else:
                    estudiante = soli.persona
                subject = 'DIRECTOR(A) SOLICITUD DE CAMBIO DE CARRERA PENDIENTE DE REVISIÓN {}'.format(
                    estudiante.__str__())
                template = 'emails/cambiocarrera_director.html'
                datos_email = {'sistema': 'SGA UNEMI', 'filtro': soli,
                               'url_boton': 'https://sga.unemi.edu.ec/alu_cambiocarrera?action=solicitantes&id=2&search=' + estudiante.cedula}
                dir_carrera = CoordinadorCarrera.objects.filter(carrera=soli.carreradestino, periodo=periodo,
                                                                tipo=3).first()
                email_director = dir_carrera.persona.emailinst if dir_carrera else ''
                lista_email = [email_director, ]
                # lista_email = ['cgomezm3@unemi.edu.ec',]
                send_html_mail(subject, template, datos_email, lista_email, [], [], cuenta=CUENTAS_CORREOS[4][1])
                response = JsonResponse({'resp': True})
            except Exception as ex:
                response = JsonResponse({'resp': False, 'mensaje': ex})
            return HttpResponse(response.content)

        elif action == 'notidecano':
            try:
                soli = SolicitudCambioCarrera.objects.get(pk=request.POST['id'])
                soli.fecha_notificacion_decano = datetime.now()
                soli.save(request)
                if soli.inscripcion:
                    estudiante = soli.inscripcion.persona
                else:
                    estudiante = soli.persona
                subject = 'DECANO(A) TIENE UNA SOLICITUD DE CAMBIO DE CARRERA PENDIENTE DE REVISIÓN {}'.format(estudiante.__str__())
                template = 'emails/cambiocarrera_director.html'
                datos_email = {'sistema': 'SGA UNEMI', 'filtro': soli,
                               'url_boton': 'https://sga.unemi.edu.ec/alu_cambiocarrera?action=solicitantes&id=' + str(
                                   soli.periodocambiocarrera.pk) + '&search=' + str(estudiante.cedula)}
                decano_facultad = soli.get_decano()
                email_decano = decano_facultad.persona.emailinst if decano_facultad else ''
                lista_email = [email_decano, ]
                # lista_email = ['cgomezm3@unemi.edu.ec',]
                send_html_mail(subject, template, datos_email, lista_email, [], [], cuenta=CUENTAS_CORREOS[4][1])
                response = JsonResponse({'resp': True})
            except Exception as ex:
                transaction.set_rollback(True)
                response = JsonResponse({'resp': False, 'mensaje': ex})
            return HttpResponse(response.content)

        elif action == 'adicionarcarrera':
            try:
                postar = AperturaPeriodoCambioCarrera.objects.get(id=int(request.POST['id']))
                carreraids = request.POST['ids'].split(',')
                for a in carreraids:
                    carrera = Carrera.objects.get(pk=a)
                    if not CarrerasCambioCarrera.objects.filter(status=True, periodocambiocarrera=postar,
                                                                carrera=carrera).exists():
                        actividad = CarrerasCambioCarrera(periodocambiocarrera=postar, carrera=carrera)
                        actividad.save(request)
                        log(u'Adiciono carrera periodo de cambio de carrera: %s %s' % (postar, carrera.nombre), request,
                            "add")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'addcupo':
            try:
                carrera = CarrerasCambioCarrera.objects.get(id=request.POST['id'])
                cupo = int(request.POST['valor'])
                if not cupo < carrera.total_solicitudes():
                    if cupo != carrera.cupo:
                        carrera.cupo = cupo
                        carrera.save(request)
                        log(u'Adiciono Cupo a carrera de Cambio de Carrera: %s' % carrera.carrera, request, "add")
                    return JsonResponse({"result": True}, safe=False)
                if cupo <= 0:
                    return JsonResponse({"result": False, 'mensaje': 'No puede ingresar un cupo menor o igual a 0'},
                                        safe=False)
                return JsonResponse({"result": False,
                                     'mensaje': 'No puede ingresar un cupo menor al total de inscritos (%s)'.format(
                                         carrera.total_solicitudes())}, safe=False)
            except Exception as e:
                transaction.set_rollback(True)
                return JsonResponse({"result": False, 'mensaje': str(e)}, safe=False)

        elif action == 'addpuntaje':
            try:
                carrera = CarrerasCambioCarrera.objects.get(id=request.POST['id'])
                solicitudes=carrera.solicitudes()
                puntaje = float(request.POST['valor'])
                if puntaje <= 0:
                    return JsonResponse({"result": False, 'mensaje': 'No puede ingresar un puntaje menor o igual a 0'},
                                        safe=False)

                if puntaje != carrera.puntajerequerido:
                    carrera.puntajerequerido = puntaje
                    carrera.save(request)
                    log(u'Adiciono Puntaje requerido a carrera de Cambio de Carrera: %s' % carrera.carrera, request,
                        "add")

                if solicitudes:
                    for solicitud in solicitudes:
                        if puntaje > float(solicitud.puntajealumno):
                            solicitud.estados=2
                            solicitud.aprobacion_admision=2
                            solicitud.observacion_admision = 'ESTIMAD@ POSTULANTE DE CAMBIO DE CARRERA / IES / MODALIDAD,' \
                                                             'SE INFORMA QUE LOS PUNTAJES MOSTRADOS AL MOMENTO DE INSCRIBIRSE ' \
                                                             'AL PROCESO FUERON TENTATIVOS Y SE ENCONTRABAN SUJETOS A CAMBIOS,' \
                                                             'POR LO TANTO SE INFORMA QUE SU REQUERIMIENTO NO PROCEDE POR NO ' \
                                                             'CUMPLIR CON EL PUNTAJE REQUERIDO DE ACUERDO A LO ESTABLECIDO' \
                                                             'EN EL ART. 96 DEL RÉGIMEN ACADÉMICO,EN CONCORDANCIA AL ART. 83' \
                                                             'DEL REGLAMENTO DEL SNNA.'
                            solicitud.save()
                            # notificacion de sga
                            asunto = u"Solicitud de cambio de carrera/IES {}".format(solicitud.get_estados_display().lower())
                            if solicitud.inscripcion:
                                para = solicitud.inscripcion.persona
                                noti = Notificacion(cuerpo=solicitud.observacion_admision, titulo=asunto,
                                                    destinatario=para,
                                                    url='/alu_solicitudcambiocarrera?action=verproceso&id={}'.format(
                                                        encrypt(solicitud.pk)), prioridad=1, app_label='SGA',
                                                    fecha_hora_visible=datetime.now() + timedelta(days=1),
                                                    tipo=2, en_proceso=False)
                                noti.save()
                            else:
                                para = solicitud.persona

                            # correo electronico para notificar a los estudiantes
                            template = 'emails/notificacion_cambio_carrera_rechazo.html'
                            datos_email = {'sistema': 'SGA UNEMI', 'puntaje': solicitud.puntajealumno,
                                           'asunto':asunto,
                                           'fecha': datetime.now().date(),
                                           'hora': datetime.now().time(),
                                           'carrera':carrera.carrera.nombre,
                                           'persona':para,
                                           'puntajerequerido':puntaje,
                                           'mensaje': solicitud.observacion_admision,
                                           'url_boton': 'sga.unemi.edu.ec/alu_solicitudcambiocarrera?action=verproceso&id={}'.format(
                                               encrypt(solicitud.pk))}
                            lista_email = para.lista_emails()
                            # lista_email = ['jguachuns@unemi.edu.ec',]
                            send_html_mail(asunto, template, datos_email, lista_email, [], [],
                                           cuenta=CUENTAS_CORREOS[4][1])

                return JsonResponse({"result": True}, safe=False)
            except Exception as e:
                return JsonResponse({"result": False, 'mensaje': str(e)}, safe=False)

        elif action == 'deletecarrera':
            try:
                carrera = CarrerasCambioCarrera.objects.get(id=request.POST['id'])
                if not carrera.con_documentos():
                    carrera.status = False
                    carrera.save(request)
                    log(u'Elimino carrera de Cambio de Carrera: %s' % carrera.carrera, request, "add")
                    return JsonResponse({"error": False})
                return JsonResponse({"error": True, 'mensaje': 'No puede eliminar esta carrera'}, safe=False)
            except Exception as e:
                transaction.set_rollback(True)
                return JsonResponse({"error": True, 'mensaje': str(e)}, safe=False)

        elif action == 'notipuntaje':
            try:
                soli = SolicitudCambioCarrera.objects.get(pk=request.POST['id'])
                soli.puntajeincorrecto = True
                soli.save(request)
                asunto = u"PUNTAJE PARA CAMBIO DE CARRERA ES INCORRECTO"

                tipocambio=True
                if not soli.persona:
                    para = soli.inscripcion.persona
                    notificacion(asunto, 'TE INFORMAMOS QUE EL PUNTAJE INGRESADO ES INCORRECTO', para, None,
                                 '/alu_solicitudcambiocarrera?action=verproceso&id={}'.format(encrypt(soli.pk)), soli.pk, 1,
                                 'sga', SolicitudCambioCarrera, request)
                else:
                    tipocambio=False
                    titulo = "Solicitud de cambio de IES pendiente de revisión"
                    lista_email = soli.persona.lista_emails(), [],
                    # lista_email = ['jguachuns@unemi.edu.ec', ]
                    datos_email = {'sistema': request.session['nombresistema'],
                                   'fecha': datetime.now().date(),
                                   'hora': datetime.now().time(),
                                   'persona': soli.persona,
                                   'tipocambio': tipocambio,
                                   'estado': soli.get_estados_display,
                                   'asunto': asunto,
                                   'mensaje': 'Te informamos que el puntaje ingresado en tu solicitud es incorrecto'}
                    template = "emails/notificacion_cambio_ies.html"
                    send_html_mail(titulo, template, datos_email, lista_email, [], [], cuenta=CUENTAS_CORREOS[0][1])

                response = JsonResponse({'resp': True})
            except Exception as ex:
                transaction.set_rollback(True)
                response = JsonResponse({'resp': False, 'mensaje': ex})
            return HttpResponse(response.content)

        elif action == 'buscarfuncionarios':
            try:
                item = []
                param = request.POST['term'].strip()
                ss = param.split(' ')
                if len(ss) == 1:
                    administrativos = Administrativo.objects.filter(Q(persona__nombres__icontains=param) |
                                                         Q(persona__apellido1__icontains=param) |
                                                         Q(persona__apellido2__icontains=param) |
                                                         Q(persona__cedula__icontains=param) |
                                                         Q(persona__pasaporte__icontains=param) |
                                                         Q(persona__usuario__username__icontains=param)).distinct()
                else:
                    administrativos = Administrativo.objects.filter(Q(persona__apellido1__icontains=ss[0]) &
                                                         Q(persona__apellido2__icontains=ss[1])).distinct()
                for a in administrativos:
                    text = str(a.persona)
                    item.append({'id': a.persona.id, 'text': text})
                return JsonResponse(item, safe=False)
            except Exception as ex:
                pass

        elif action == 'addresponsable':
            try:
                idpersona=int(request.POST['idpersona'])
                idrol=int(request.POST['idrol'])
                estado=eval(request.POST['estado'])
                responsable = ReponsableCambioCarrera(persona_id=idpersona, rol=idrol, estado=estado)
                responsable.save(request)
                log(u'Adiciona funcionarios responsables de cambio de carrera: %s - %s - %s' , request, "addresponsable")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": "Error al guardar los datos."})

        elif action == 'editresponsable':
            try:
                responsable = ReponsableCambioCarrera.objects.get(pk=int(request.POST['id']))
                responsable.persona_id = int(request.POST['idpersona'])
                responsable.rol= int(request.POST['idrol'])
                responsable.estado = eval(request.POST['estado'])
                responsable.save(request)
                log(u'Edito responsable cambio carrera: %s - %s - %s', request, "edit")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": "Error al guardar los datos."})

        elif action == 'delresponsable':
            try:
                with transaction.atomic():
                    responsable = ReponsableCambioCarrera.objects.get(pk=int(request.POST['id']))
                    responsable.status = False
                    responsable.save(request)
                    log(u'Elimino registro de responsable de cambio de carrera: %s - %s - %s', request, "del")
                    res_json = {"error": False}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        elif action == 'reasignarsolicitud':
            try:
                solicitud=SolicitudCambioCarrera.objects.get(id=request.POST['idsolicitud'])
                solicitud.persona_admision_id=request.POST['idfuncionario']
                solicitud.save(request)
                log(u'Reasigno funcionario para revision: %s - %s - %s', request, "edit")
                return JsonResponse({"result": True, "mensaje": "Registro Guardado"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": False, "mensaje": "Error al reasignar funcionario."})

        if action=='addevidenciavalidada':
            try:
                id=request.POST['id']
                # observacion=request.POST['observacion']
                solicitud=SolicitudCambioCarrera.objects.get(id=id)
                if solicitud.doc_evidencia_validados():
                    for doc in solicitud.doc_evidencia_validados():
                        doc.status=False
                        doc.save(request)
                if request.FILES:
                    documentos=request.FILES.getlist('archivos[]')
                    for documento in documentos:
                        extension = documento._name.split('.')
                        tam = len(extension)
                        exte = extension[tam - 1]
                        if documento.size > 2194304:
                            # return JsonResponse({"result": True, "mensaje": u"Error, el tamaño del archivo es mayor a 2 Mb."})
                            raise NameError('Error, el tamaño del archivo es mayor a 2 Mb.')
                        if not exte in ['pdf']:
                            # return JsonResponse({"result": True, "mensaje": u"Error, solo archivos .pdf"})
                            raise NameError('Error, solo archivos .pdf')
                        nombre_persona = remover_caracteres_especiales_unicode(solicitud.persona.apellido1).lower().replace(' ','_')
                        nombredoc = "doc_evidencia"
                        documento._name = generar_nombre("{}__{}".format(nombredoc,nombre_persona),documento._name)
                        docvalidado = DocumentosValidados(solicitud_id=id, archivo=documento)
                        docvalidado.save(request)
                        log(u'Adiciona Archivo de veracidad de requisitos: %s - %s - %s', request,"addevidenciavalidada")
                else:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": True, "mensaje": u"Seleccione archivos con formato pdf."})
                return JsonResponse({"result": False})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": u"Error al obtener los datos. {}".format(ex)})

        elif action == 'notiestudiante':
            try:
                if not request.POST['texto']:
                    response = JsonResponse({'result': False})
                    return HttpResponse(response.content)
                soli = SolicitudCambioCarrera.objects.get(pk=request.POST['id'])
                soli.mensajedocumentosincompletos=request.POST['texto']
                soli.estadomensaje=True
                soli.save(request)
                asunto = u"DOCUMENTOS DE SOLICITUD DE CAMBIO DE CARRERA"
                tipocambio = True
                if not soli.persona:
                    para = soli.inscripcion.persona
                    notificacion(asunto, request.POST['texto'], para, None,
                                 '/alu_solicitudcambiocarrera?action=verproceso&id={}'.format(encrypt(soli.pk)), soli.pk, 1,
                                 'sga', SolicitudCambioCarrera, request)
                else:
                    tipocambio = False
                    para = soli.persona
                titulo = "Solicitud de cambio de carrera/IES pendiente de revisión"
                lista_email = para.lista_emails()
                # lista_email = ['jguachuns@unemi.edu.ec', ]
                datos_email = {'sistema': request.session['nombresistema'],
                               'fecha': datetime.now().date(),
                               'hora': datetime.now().time(),
                               'tipocambio':tipocambio,
                               'estado':soli.get_estados_display,
                               'persona': para,
                               'asunto': asunto,
                               'mensaje': request.POST['texto']}
                template = "emails/notificacion_cambio_ies.html"
                send_html_mail(titulo, template, datos_email, lista_email, [], [], cuenta=CUENTAS_CORREOS[0][1])

                response = JsonResponse({'result': True})
            except Exception as ex:
                transaction.set_rollback(True)
                response = JsonResponse({'result': False, 'mensaje': ex})
            return HttpResponse(response.content)

        elif action == 'rechazarsolicitud':
            try:
                with transaction.atomic():
                    if not request.POST['observacion']:
                        response = JsonResponse({'result': False,'mensaje':'Por Favor escriba el motivo del rechazo de solicitud'})
                        return HttpResponse(response.content)
                    filtro = SolicitudCambioCarrera.objects.get(pk=int(request.POST['id']))
                    filtro.revision_admision=2
                    filtro.aprobacion_admision=2
                    filtro.estados=6
                    filtro.observacion_admision = request.POST['observacion'].capitalize()
                    filtro.save(request)

                    asunto = u"VALIDACIÓN DE DOCUMENTOS DE CAMBIO DE CARRERA/IES {}".format(filtro.get_estados_display())
                    tipocambio = True
                    if not filtro.persona:
                        para = filtro.inscripcion.persona
                        notificacion(asunto, request.POST['observacion'], para, None,
                                     '/alu_solicitudcambiocarrera?action=verproceso&id={}'.format(encrypt(filtro.pk)),
                                     filtro.pk, 1,
                                     'sga', SolicitudCambioCarrera, request)
                    else:
                        tipocambio = False
                        para = filtro.persona
                    titulo = u"Validación de documentos de cambio de carrera/IES {}".format(filtro.get_estados_display().lower())
                    lista_email = para.lista_emails()
                    # lista_email = ['jguachuns@unemi.edu.ec', ]
                    datos_email = {'sistema': request.session['nombresistema'],
                                   'fecha': datetime.now().date(),
                                   'hora': datetime.now().time(),
                                   'persona': para,
                                   'tipocambio':tipocambio,
                                   'estado':filtro.get_estados_display,
                                   'asunto': asunto,
                                   'mensaje': request.POST['observacion'].capitalize()}
                    template = "emails/notificacion_cambio_ies.html"
                    send_html_mail(titulo, template, datos_email, lista_email, [], [], cuenta=CUENTAS_CORREOS[0][1])
                    log(u'Departamento de Admision rechazo solicitud: {} {}'.format(
                        filtro.inscripcion, filtro.get_revision_admision_display()), request, "edit")
                    response = JsonResponse({'result': True})
            except Exception as ex:
                transaction.set_rollback(True)
                response = JsonResponse({'result': False, 'mensaje': ex})
            return HttpResponse(response.content)

        elif action == 'permitirsolicitud':
            try:
                with transaction.atomic():
                    filtro = SolicitudCambioCarrera.objects.get(pk=int(request.POST['id']))
                    filtro.revision_admision=filtro.documentos_aprobados()
                    filtro.aprobacion_admision=0
                    filtro.estados=0
                    filtro.save(request)
                    # asunto = u"VALIDACIÓN DE DOCUMENTOS DE CAMBIO DE CARRERA/IES {}".format(filtro.get_estados_display())
                    # if not filtro.persona:
                    #     para = filtro.inscripcion.persona
                    #     notificacion(asunto, request.POST['observacion'], para, None,
                    #                  '/alu_solicitudcambiocarrera?action=verproceso&id={}'.format(encrypt(filtro.pk)),
                    #                  filtro.pk, 1,
                    #                  'sga', SolicitudCambioCarrera, request)
                    # else:
                    #     para = filtro.persona
                    # titulo = asunto
                    # lista_email = para.lista_emails()
                    # # lista_email = ['jguachuns@unemi.edu.ec', ]
                    # datos_email = {'sistema': request.session['nombresistema'],
                    #                'fecha': datetime.now().date(),
                    #                'hora': datetime.now().time(),
                    #                'persona': para,
                    #                'asunto': asunto,
                    #                'mensaje': request.POST['observacion']}
                    # template = "emails/notificacion_cambio_ies.html"
                    # send_html_mail(titulo, template, datos_email, lista_email, [], [], cuenta=CUENTAS_CORREOS[0][1])

                    log(u'Departamento de Admision permitio solicitud {} {}'.format(
                        filtro.inscripcion, filtro.get_revision_admision_display()), request, "edit")
                    response = JsonResponse({'result': True})
            except Exception as ex:
                transaction.set_rollback(True)
                response = JsonResponse({'result': False, 'mensaje': ex})
            return HttpResponse(response.content)

        elif action == 'subiraprobacionadmision':
            try:
                if 'id' not in request.POST:
                    return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al procesar la solicitud", "showSwal": "True", "swalType": "error"})

                archivo = request.FILES['archivo']
                descripcionarchivo = 'Archivo de Aprobación'

                # Validar el archivo
                resp = validar_archivo(descripcionarchivo, archivo, ['PDF'], '4MB')
                if resp['estado'] != "OK":
                    return JsonResponse({"result": "bad", "titulo": "Atención!!!", "mensaje": resp["mensaje"], "showSwal": "True", "swalType": "warning"})

                # Consulto la solicitud
                solicitud = SolicitudCambioCarrera.objects.get(pk=int(encrypt(request.POST['id'])))

                apellido1 = solicitud.inscripcion.persona.apellido1 if solicitud.inscripcion else solicitud.persona.apellido1
                nombre_persona = remover_caracteres_especiales_unicode(apellido1).lower().replace(' ', '_')

                archivo._name = generar_nombre(nombre_persona + "INFORME_ADMISION_CAMBIO_CARRERA", archivo._name)

                solicitud.archivoinformeadmsion = archivo
                solicitud.save(request)

                log(u'%s actualizó archivo de aprobación de admisión para solicitud: %s' % (persona, solicitud), request, "edit")
                return JsonResponse({"result": "ok", "titulo": "Proceso exitoso!!!", "mensaje": u"Registro actualizado con éxito", "showSwal": True, "idsol": encrypt(solicitud.id)})
            except Exception as ex:
                msg = ex.__str__()
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al guardar los datos. [%s]" % msg, "showSwal": "True", "swalType": "error"})

        elif action == 'deldocumento':
            try:
                with transaction.atomic():
                    documento = DocumentosSolicitudCambioCarrera.objects.get(pk=int(request.POST['id']))
                    documento.motivo=request.POST['observacion']
                    documento.status = False
                    documento.save(request)
                    log(u'Elimino documento de estudiante de cambio de carrera: %s - %s - %s', request, "deldocumento")
                    res_json = {"error": False}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        elif action == 'actualizarestado':
            try:
                with transaction.atomic():
                    solicitud = SolicitudCambioCarrera.objects.get(pk=int(request.POST['id']))
                    if solicitud.documentos_aprobados() == 1 and solicitud.revision_admision != 1:
                        if solicitud.revision_bienestar == 0:
                            verificador = SolicitudCambioCarrera.objects.filter(status=True, periodocambiocarrera=solicitud.periodocambiocarrera).values('id')
                            cont = 0
                            persona_bienestar_id = 0
                            responsables = ReponsableCambioCarrera.objects.filter(status=True, rol=1,estado=True).values_list('persona_id',flat=True)
                            resinactivos = ReponsableCambioCarrera.objects.filter(status=True, rol=1,estado=False).values_list('persona_id',flat=True)
                            if not responsables:
                                transaction.set_rollback(True)
                                return JsonResponse({"result": "bad",
                                                     "mensaje": u"No existen Funcionarios de Bienestar configurados para asignar solicitud, configure los funcionarios ."})
                            # lista = []
                            for p in responsables:
                                cont += 1
                                if not verificador.filter(persona_bienestar=p).exists():
                                    persona_bienestar_id = p
                                    break
                            if persona_bienestar_id == 0 and verificador.exists():
                                persona_bienestar_id = SolicitudCambioCarrera.objects.filter(status=True,periodocambiocarrera=solicitud.periodocambiocarrera).exclude(persona_bienestar_id__in=resinactivos).values('persona_bienestar_id').annotate(total=Count('id')).order_by('total')[0]['persona_bienestar_id']
                            solicitud.persona_bienestar_id = int(persona_bienestar_id)
                        solicitud.revision_admision = 1
                        solicitud.save()
                        notificacion('SOLICITUD DE CAMBIO DE CARRERA PENDIENTE DE REVISIÓN',
                                     'Se le ha asignado una solicitud de cambio de carrera para revisión',
                                     solicitud.persona_bienestar, None,
                                     '/alu_cambiocarrera?action=solicitantes&id={}&search={}'.format(
                                         solicitud.periodocambiocarrera.pk,
                                         solicitud.inscripcion.persona.cedula if solicitud.inscripcion else solicitud.persona.cedula),
                                     solicitud.pk, 1,
                                     'sga', SolicitudCambioCarrera, request)
                        log(u'Acualizo estado de revision: {} {}'.format(solicitud.inscripcion, solicitud.get_revision_admision_display()), request, "actualizarestado")
                    response = JsonResponse({'result': True})
            except Exception as ex:
                transaction.set_rollback(True)
                response = JsonResponse({'result': False, 'mensaje': ex})
            return HttpResponse(response.content)
        return JsonResponse({"result": "bad", "mensaje": u"Solicitud Incorrecta."})
    else:
        # fin get
        adduserdata(request, data)
        if 'action' in request.GET:
            data['action'] = action = request.GET['action']

            if action == 'documentosrequeridos':
                try:
                    data['title'] = u'Documentos Requeridos'
                    data['evento'] = RequisitosCambioCarrera.objects.filter(status=True).order_by('fecha_creacion')
                    return render(request, "alu_cambiocarrera/viewdocumento.html", data)
                except Exception as ex:
                    pass

            elif action == 'adddocumento':
                try:
                    data['form2'] = DocumentoRequeridoCambioCarreraForm()
                    template = get_template("alu_practicaspreprofesionalesinscripcion/modal/formdocumentos.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'editdocumento':
                try:
                    data['id'] = request.GET['id']
                    data['filtro'] = filtro = RequisitosCambioCarrera.objects.get(pk=request.GET['id'])
                    data['form2'] = DocumentoRequeridoCambioCarreraForm(initial=model_to_dict(filtro))
                    template = get_template("alu_practicaspreprofesionalesinscripcion/modal/formdocumentos.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'addaperturaperiodo':
                try:
                    data['title'] = u'Adicionar periodo de solicitud para cambio de carrera'
                    form = AperturaPeriodoCambioCarreraForm(initial={'periodo': periodo})
                    form.adicionar()
                    data['form'] = form
                    return render(request, "alu_cambiocarrera/addaperturaperiodo.html", data)
                except Exception as ex:
                    pass

            elif action == 'editaperturaperiodo':
                try:
                    data['title'] = u'Editar apertura de Periodo para cambio de carrera'
                    data['apertura'] = apertura = AperturaPeriodoCambioCarrera.objects.get(pk=int(request.GET['id']))
                    form = AperturaPeriodoCambioCarreraForm(initial=model_to_dict(apertura))
                    data['form'] = form
                    return render(request,
                                  "alu_cambiocarrera/addaperturaperiodo.html", data)
                except Exception as ex:
                    pass

            elif action == 'delaperturasolicitud':
                try:
                    data['title'] = u'Eliminar periodo de cambio de carrera'
                    data['periodocambio'] = AperturaPeriodoCambioCarrera.objects.get(pk=int(request.GET['id']))
                    return render(request, "alu_cambiocarrera/delperiodocambiocarrera.html", data)
                except Exception as ex:
                    pass

            elif action == 'solicitantes':
                try:
                    data['title'] = u'Solicitantes de Cambio de Carrera'
                    data['id'] = id = int(request.GET['id'])
                    data['apertura'] = apertura = AperturaPeriodoCambioCarrera.objects.get(pk=id)
                    data['estado_solicitud'] = ESTADO_SOLICITUD_CAMBIO_CARRERA
                    data['estado_pasos'] = ESTADOS_PASOS_SOLICITUD
                    valcarreras = False
                    querybase = SolicitudCambioCarrera.objects.filter(periodocambiocarrera=apertura, status=True)
                    if es_decano:
                        if len(querydecano)>1:
                            coordinaciones=[]
                            for decano in querydecano:
                                coordinaciones.append(decano.coordinacion)
                            coordinacion=Q(carreradestino__coordinacion__in=coordinaciones)
                        else:
                            coordinacion=Q(carreradestino__coordinacion=querydecano.last().coordinacion)
                        querybase = querybase.filter(coordinacion)
                    carreras = querybase.values_list('carreradestino', 'carreradestino__nombre').distinct()
                    if es_director_carr:
                        carreras = carreras.filter(carreradestino__in=miscarreras.values_list('id', flat=True))
                        valcarreras = True
                    if es_asist_bienestar:
                        querybase = querybase.filter(persona_bienestar=persona, revision_admision=1)

                    data['carreras'] = carreras
                    data['funcionarios']=personas_admision
                    # if not es_decano and not es_director_carr:
                    #     querybase=querybase.filter(persona_admision_id=persona.id)

                    estsolicitud, carrera, desde, hasta, search, filtros, url_vars, funcionario = request.GET.get('estsolicitud',
                                                                                                     ''), request.GET.get(
                        'carrera', ''), request.GET.get('desde', ''), request.GET.get('hasta', ''), request.GET.get(
                        'search', ''), Q(status=True), '', request.GET.get('funcionario', ''),

                    if personas_admision.filter(persona=persona).exists() and funcionario == '':
                        filtros = filtros & Q(persona_admision_id=persona)
                        data['funcionario'] = persona.id

                    # if es_director_adm:
                    #     filtros = filtros & Q(revision_bienestar=1)

                    if estsolicitud:
                        data['estsolicitud'] = estsolicitud = int(estsolicitud)
                        url_vars += "&estsolicitud={}".format(estsolicitud)
                        filtros = filtros & Q(estados=estsolicitud)
                    if desde:
                        data['desde'] = desde
                        url_vars += "&desde={}".format(desde)
                        filtros = filtros & Q(fecha_creacion__gte=desde)
                    if hasta:
                        data['hasta'] = hasta
                        url_vars += "&hasta={}".format(hasta)
                        filtros = filtros & Q(fecha_creacion__lte=hasta)
                    if carrera:
                        data['carrera'] = int(carrera)
                        url_vars += "&carrera={}".format(carrera)
                        filtros = filtros & Q(carreradestino_id=carrera)

                    if funcionario and funcionario != '0':
                        data['funcionario']=int(funcionario)
                        url_vars += "&funcionario={}".format(funcionario)
                        filtros = filtros & Q(persona_admision_id=funcionario)

                    if 'idsol' in request.GET:
                        filtros = filtros & Q(pk=int(encrypt(request.GET['idsol'])))

                    if search:
                        data['search'] = search
                        s = search.split()
                        if len(s) == 1:
                            filtros = filtros & (Q(inscripcion__persona__apellido2__icontains=search) | Q(
                                inscripcion__persona__cedula__icontains=search) | Q(
                                inscripcion__persona__apellido1__icontains=search) |
                                                 Q(persona__apellido2__icontains=search) | Q(
                                        persona__cedula__icontains=search) | Q(
                                        persona__apellido1__icontains=search))
                        else:
                            filtros = filtros & (Q(inscripcion__persona__apellido1__icontains=s[0]) & Q(
                                inscripcion__persona__apellido2__icontains=s[1])
                                                 | Q(persona__apellido1__icontains=s[0]) & Q(
                                        persona__apellido2__icontains=s[1]))
                        url_vars += '&search={}'.format(search)
                    url_vars += '&action={}&id={}'.format(action, id)
                    data["url_vars"] = url_vars
                    if valcarreras:
                        query = querybase.filter(filtros).select_related('inscripcion').filter(
                            carreradestino__id__in=miscarreras.values_list('id', flat=True)).order_by('pk')
                    else:
                        query = querybase.select_related('inscripcion').filter(filtros).order_by('pk')
                    data['listcount'] = query.count()
                    paging = MiPaginador(query, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['lista'] = page.object_list
                    return render(request, "alu_cambiocarrera/solicitantescambiocarrera.html", data)
                except Exception as ex:
                    pass

            elif action == 'descargarexpediente':
                try:
                    dominiosistema = request.build_absolute_uri('/')[:-1].strip("/")
                    solicitud = int(request.GET['id'])
                    informes = DocumentosSolicitudCambioCarrera.objects.filter(status=True, solicitud=solicitud).distinct()
                    archivos_lista = []
                    directory = os.path.join(SITE_STORAGE, 'CambioCarreraAlumno')

                    try:
                        os.stat(directory)
                    except:
                        os.mkdir(directory)

                    url = os.path.join(SITE_STORAGE, 'media', 'CambioCarreraAlumno',
                                       'expedientesolicitante_{}_{}.zip'.format(solicitud,
                                                                               random.randint(1, 10000).__str__()))
                    url_zip = url
                    fantasy_zip = zipfile.ZipFile(url, 'w')
                    solicitante=''
                    cont=0
                    for inf in informes:
                        cont+=1
                        if inf.archivo:
                            if inf.solicitud.inscripcion:
                                estudiante=inf.solicitud.inscripcion.persona
                            else:
                                estudiante=inf.solicitud.persona
                            solicitante = remover_caracteres_especiales_unicode(
                                estudiante.__str__().lower().replace(' ', '_')).lower().replace(' ', '_')
                            doc = remover_caracteres_especiales_unicode(
                                inf.documento.nombre.__str__().lower().replace(' ', '_')).lower().replace(' ', '_')
                            fantasy_zip.write(inf.archivo.path,
                                              '{}_{}_{}.pdf'.format(cont, solicitante, doc))
                    fantasy_zip.close()
                    response = HttpResponse(open(url_zip, 'rb'), content_type='application/zip')
                    response['Content-Disposition'] = 'attachment; filename=expedientesolicitante_{}_{}.zip'.format(solicitante,
                                                                                                                   random.randint(1,10000).__str__())
                    return response
                except Exception as ex:
                    print(ex)
                    print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
                    messages.error(request,solicitante )
                    return redirect('{}?action=solicitantes&id={}&linea={}'.format(request.path, ex,
                                                                                     sys.exc_info()[-1].tb_lineno))

            elif action == 'descargarexpedientecompleto':
                try:
                    dominiosistema = request.build_absolute_uri('/')[:-1].strip("/")
                    id=int(request.GET['id'])
                    solicitud = SolicitudCambioCarrera.objects.get(pk=id)
                    informes = solicitud.documentoscargados()
                    evidencias = solicitud.doc_evidencia_validados()
                    if solicitud.inscripcion:
                        estudiante = solicitud.inscripcion.persona
                    else:
                        estudiante = solicitud.persona

                    archivos_lista = []
                    directory = os.path.join(SITE_STORAGE, 'CambioCarreraAlumno')

                    try:
                        os.stat(directory)
                    except:
                        os.mkdir(directory)

                    url = os.path.join(SITE_STORAGE, 'media', 'CambioCarreraAlumno',
                                       'expedientesolicitante_{}_{}.zip'.format(id, random.randint(1, 10000).__str__()))
                    url_zip = url
                    fantasy_zip = zipfile.ZipFile(url, 'w')
                    solicitante = ''
                    cont = 0
                    for inf in informes:
                        cont += 1
                        if inf.archivo:
                            solicitante = remover_caracteres_especiales_unicode(
                                estudiante.__str__().lower().replace(' ', '_')).lower().replace(' ', '_')
                            doc = remover_caracteres_especiales_unicode(
                                inf.documento.nombre.__str__().lower().replace(' ', '_')).lower().replace(' ', '_')
                            fantasy_zip.write(inf.archivo.path,'{}_{}_{}.pdf'.format(cont, solicitante, doc))
                    cont=0
                    for evi in evidencias:
                        if evi.archivo:
                            solicitante = remover_caracteres_especiales_unicode(
                                estudiante.__str__().lower().replace(' ', '_')).lower().replace(' ', '_')
                            # doc = generar_nombre("{}_{}".format("evidencia_verificada","verificada"),evi.archivo._name)
                            fantasy_zip.write(evi.archivo.path, 'evidencia_verificada_{}_{}.pdf'.format(solicitante, cont))
                            cont+=1
                    fantasy_zip.write(solicitud.archivobienestar.path, 'informe_bienestar.pdf')
                    fantasy_zip.close()
                    response = HttpResponse(open(url_zip, 'rb'), content_type='application/zip')
                    response['Content-Disposition'] = 'attachment; filename=expedientecompletosolicitante_{}_{}.zip'.format(
                        solicitante,
                        random.randint(1, 10000).__str__())
                    return response
                except Exception as ex:
                    print(ex)
                    print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
                    messages.error(request, solicitante)
                    return redirect('{}?action=solicitantes&id={}&linea={}'.format(request.path, ex,
                                                                                   sys.exc_info()[-1].tb_lineno))

            elif action == 'descargarrevision':
                try:
                    __author__ = 'Unemi'
                    title = easyxf('font: name Times New Roman, color-index black, bold on , height 350; alignment: horiz centre')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    style1 = easyxf(num_format_str='D-MMM-YY')
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('Historial-Correciones')
                    ws.write_merge(0, 0, 0, 5, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=Proceso de revision- ' + '-' + random.randint(1, 10000).__str__() + '.xls'
                    columns = [
                        (u"N°", 3000),
                        (u"SOLICITANTE", 10000),
                        (u"FECHA REVISION", 10000),
                        (u"ESTADO", 6000),
                        (u"QUIEN REVISO", 10000),
                        (u"OBSERVACION", 15000),
                    ]
                    row_num = 1
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    seguimiento= SeguimientoRevisionDocumentoCC.objects.filter(solictud=request.GET['id'])
                    mensaje='NO REGISTRA'
                    row_num = 2
                    for s in seguimiento:
                        ws.write(row_num, 0, s.id, font_style2)
                        ws.write(row_num, 1, str(s.solictud.inscripcion if s.solictud.inscripcion else mensaje), font_style2)
                        ws.write(row_num, 2, s.fecha if s.fecha else mensaje, style1)
                        ws.write(row_num, 3, s.get_estados_display(), font_style2)
                        ws.write(row_num, 4, str(s.revisor if s.revisor else mensaje), font_style2)
                        ws.write(row_num, 5, s.observacion if s.observacion else mensaje, font_style2)
                        row_num += 1
                    wb.save(response)
                    return response
                except Exception as ex:
                    pass

            elif action == 'descargarreporte':
                try:
                    __author__ = 'Unemi'
                    title = easyxf('font: name Times New Roman, color-index black, bold on , height 350; alignment: horiz centre')
                    title2 = easyxf('font: name Arial, color-index black, bold on , height 200; alignment: horiz centre')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    style1 = easyxf(num_format_str='D-MMM-YY')
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('Reporte de Solicitudes')
                    ws.write_merge(0, 0, 0, 13, 'REPORTE DE CAMBIO DE CARRERAS/IES/MODALIDAD', title)
                    ws.write_merge(1, 1, 3, 5, 'INFORMACIÓN DE PROCEDENCIA', title2)
                    ws.write_merge(1, 1, 8, 9, 'UNEMI RECEPTA', title2)
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=Reporte de solicitudes cambio carrera/ies- ' + '-' + random.randint(1, 10000).__str__() + '.xls'

                    columns = [
                        (u"N°", 2000),
                        (u"NOMBRES Y APELLIDOS", 9000),
                        (u"CÉDULA", 4000),
                        (u"IES", 7000),
                        (u"FACULTAD", 10000),
                        (u"CARRERA", 10000),
                        (u"PUNTAJE COHORTE SOLICITADA", 2000),
                        (u"PUNTAJE ESTUDIANTE", 2000),
                        (u"FACULTAD", 15000),
                        (u"CARRERA", 15000),
                        (u"ATENDIDO POR", 15000),
                        (u"NÚMERO DE REVISIONES", 2000),
                        (u"ESTADO", 3000),
                    ]
                    row_num = 2
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    solicitudes= SolicitudCambioCarrera.objects.filter(status=True)
                    mensaje='NO REGISTRA'
                    row_num = 3
                    numero=0
                    for s in solicitudes:
                        seguimiento = SeguimientoRevisionDocumentoCC.objects.filter(solictud=s)
                        if s.inscripcion:
                            estudiante=s.inscripcion.persona
                            ies='UNIVERSIDAD ESTATAL DE MILAGRO'
                            carrera=s.inscripcion.carrera
                            facultad=s.inscripcion.coordinacion
                        else:
                            estudiante=s.persona
                            ies=s.universidad
                            facultad='NO REGISTRA'
                            carrera='NO REGISTRA'
                            if s.universidadtext:
                                ies=s.universidadtext
                        numero+=1
                        ws.write(row_num, 0, numero, font_style2)
                        ws.write(row_num, 1, str(estudiante), font_style2)
                        ws.write(row_num, 2, estudiante.cedula, font_style2 )
                        ws.write(row_num, 3, str(ies).upper(), font_style2)
                        ws.write(row_num, 4, str(facultad), font_style2)
                        ws.write(row_num, 5, str(carrera), font_style2)
                        ws.write(row_num, 6, s.puntaje_minimo_carrera().puntajerequerido, font_style2)
                        ws.write(row_num, 7, s.puntajealumno, font_style2)
                        ws.write(row_num, 8, str(s.carreradestino.mi_coordinacion()), font_style2)
                        ws.write(row_num, 9, str(s.carreradestino), font_style2)
                        ws.write(row_num, 10, str(s.persona_admision), font_style2)
                        ws.write(row_num, 11, len(seguimiento), font_style2)
                        ws.write(row_num, 12, s.get_estados_display(), font_style2)
                        row_num += 1
                    wb.save(response)
                    return response
                except Exception as ex:
                    pass

            elif action == 'descargarreport':
                try:
                    wb = openxl.Workbook()
                    wb["Sheet"].title = "Reporte_Solicitudes"
                    ws = wb.active
                    style_title = openxlFont(name='Arial', size=16, bold=True)
                    style_cab=openxlFont(name='Arial', size=10, bold=True)
                    alinear=alin(horizontal="center", vertical="center")
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=Reporte de solicitudes cambio carrera/ies- ' + '-' + random.randint(
                        1, 10000).__str__() + '.xlsx'

                    ws.column_dimensions['B'].width = 25
                    ws.column_dimensions['C'].width = 15
                    ws.column_dimensions['D'].width = 25
                    ws.column_dimensions['E'].width = 25
                    ws.column_dimensions['F'].width = 25
                    ws.column_dimensions['I'].width = 25
                    ws.column_dimensions['J'].width = 25
                    ws.column_dimensions['L'].width = 25
                    ws.column_dimensions['M'].width = 25
                    ws.merge_cells('A1:M1')
                    ws['A1']='REPORTE DE CAMBIO DE CARRERAS/IES/MODALIDAD'
                    celda1=ws['A1']
                    celda1.font = style_title
                    celda1.alignment = alinear

                    ws.merge_cells('D2:F2')
                    ws['D2']='INFORMACIÓN DE PROCEDENCIA'
                    ws.merge_cells('I2:J2')
                    celda2=ws['D2']
                    celda2.font = style_cab
                    celda2.alignment = alinear

                    ws['I2'] = 'UNEMI RECEPTA'
                    celda3 = ws['I2']
                    celda3.font = style_cab
                    celda3.alignment = alinear
                    columns = [u"N°",u"NOMBRES Y APELLIDOS", u"CÉDULA",
                               u"IES", u"FACULTAD", u"CARRERA", u"PUNTAJE COHORTE SOLICITADA",
                               u"PUNTAJE ESTUDIANTE",u"FACULTAD", u"CARRERA",u"NÚMERO DE REVISIONES",u"ESTADO",u"FUNCIONARIO"
                    ]
                    row_num=3
                    for col_num in range(0, len(columns)):
                        celda=ws.cell(row=row_num, column=(col_num+1), value=columns[col_num])
                        celda.font=style_cab
                    apertura=AperturaPeriodoCambioCarrera.objects.get(status=True,id=int(request.GET['id']))
                    solicitudes = SolicitudCambioCarrera.objects.filter(status=True, periodocambiocarrera=apertura)
                    mensaje = 'NO REGISTRA'
                    row_num = 4
                    cambiocarrera=0
                    cambiouniversidad=0
                    numero=0
                    for s in solicitudes:
                        seguimiento = SeguimientoRevisionDocumentoCC.objects.filter(solictud=s)
                        if s.inscripcion:
                            estudiante = s.inscripcion.persona
                            ies = 'UNIVERSIDAD ESTATAL DE MILAGRO'
                            carrera = s.inscripcion.carrera
                            facultad = s.inscripcion.coordinacion
                            cambiocarrera += 1
                        else:
                            cambiouniversidad += 1
                            estudiante = s.persona
                            ies = s.universidad
                            facultad = 'NO REGISTRA'
                            carrera = 'NO REGISTRA'
                            if s.universidadtext:
                                ies = s.universidadtext
                        numero+=1
                        ws.cell(row=row_num, column=1, value=numero)
                        ws.cell(row=row_num, column=2, value=str(estudiante))
                        ws.cell(row=row_num, column=3, value=estudiante.cedula)
                        ws.cell(row=row_num, column=4, value=str(ies).upper())
                        ws.cell(row=row_num, column=5, value=str(facultad))
                        ws.cell(row=row_num, column=6, value=str(carrera))
                        ws.cell(row=row_num, column=7, value=s.puntaje_minimo_carrera().puntajerequerido)
                        ws.cell(row=row_num, column=8, value=s.puntajealumno)
                        ws.cell(row=row_num, column=9, value=str(s.carreradestino.mi_coordinacion()))
                        ws.cell(row=row_num, column=10, value=str(s.carreradestino))
                        ws.cell(row=row_num, column=11, value=len(seguimiento))
                        ws.cell(row=row_num, column=12, value=s.get_estados_display())
                        ws.cell(row=row_num, column=13, value=str(s.persona_admision))
                        row_num += 1
                    #Graficas Estadisticas
                    ws2 = wb.create_sheet("Gráficos_Estadísticos")
                    ws2.column_dimensions['A'].width = 25
                    ws2.column_dimensions['B'].width = 20
                    ws2.column_dimensions['C'].width = 20
                    ws2.column_dimensions['H'].width = 25
                    ws2.column_dimensions['I'].width = 25
                    ws2.column_dimensions['J'].width = 25
                    ws2.merge_cells('A1:O1')
                    ws2['A1'] = 'ESTADÍTICAS DE CAMBIO DE CARRERAS/IES/MODALIDAD'
                    celdaA1 = ws2['A1']
                    celdaA1.font = style_title
                    celdaA1.alignment = alinear
                    ws2['A3'] = 'VARIABLES'
                    ws2['A3'].font = style_cab
                    ws2['B3'] = 'CANTIDAD'
                    ws2['B3'].font = style_cab

                    ws2['A4'] = 'CAMBIO DE IES'
                    ws2['A5'] = 'CAMBIO DE CARRERA'
                    ws2['B4'] = cambiouniversidad
                    ws2['B5'] = cambiocarrera
                    chart = PieChart()
                    chart.title = " GRÁFICO ESTADÍSTICO"
                    labels = Reference(ws2, min_col=1, min_row=4, max_row=5)

                    data = Reference(ws2, min_col=2, min_row=3, max_row=5)

                    chart.add_data(data, titles_from_data=True)

                    chart.set_categories(labels)

                    chart.title = "MARGEN DE SOLICITUDES"
                    ws2.add_chart(chart, "E3")

                    filas2 = [ u'PENDIENTE'
                                , u'FINALIZADO'
                                , u'RECHAZADO'
                                , u'APROBADO POR ADMISION'
                                , u'APROBADO POR DIRECTOR'
                                , u'RECHAZADO POR DIRECTOR'
                                , u'RECHAZADO POR ADMISION'
                                , u'APROBADO POR DECANO'
                                , u'RECHAZADO POR DECANO'
                               ]
                    ws2['A19'] = 'ESTADOS'
                    ws2['A19'].font = style_cab
                    ws2['B19'] = 'CAMBIO CARRERA'
                    ws2['B19'].font = style_cab
                    ws2['C19'] = 'CAMBIO IES'
                    ws2['C19'].font = style_cab
                    for fila in range(0, len(filas2)):
                        ws2.cell(row=(fila + 20), column=1, value=filas2[fila])
                        ws2.cell(row=(fila + 20), column=2, value=len(solicitudes.filter(estados=fila, inscripcion__isnull=False)))
                        ws2.cell(row=(fila + 20), column=3,value=len(solicitudes.filter(estados=fila, inscripcion__isnull=True)))

                    c1 = BarChart()
                    c1.title = 'CAMBIO DE CARRERA/IES'
                    c1.type = "bar"
                    c1.style = 10
                    c1.y_axis.title = "Número de solicitudes"
                    c1.x_axis.title = "Estados"
                    yvalues = Reference(ws2, min_col=2, min_row=19, max_col=3, max_row=(len(filas2)+20))
                    xvalues = Reference(ws2, min_col=1, min_row=20, max_row=(len(filas2)+20))

                    c1.add_data(yvalues, titles_from_data=True)
                    c1.set_categories(xvalues)
                    ws2.add_chart(c1, 'E19')

                    ws2['A35'] = 'CARRERAS'
                    ws2['A35'].font = style_cab
                    ws2['B35'] = 'CAMBIO CARRERA APROBADOS'
                    ws2['B35'].font = style_cab
                    ws2['C35'] = 'CAMBIO IES APROBADOS'
                    ws2['C35'].font = style_cab
                    carreras=apertura.carreras_periodo()
                    for idx, cambiocarrera in enumerate(carreras):
                        ws2.cell(row=(idx + 36), column=1, value=str(cambiocarrera.carrera.nombre))
                        ws2.cell(row=(idx + 36), column=2, value=len(solicitudes.filter(revision_director=1, carreradestino=cambiocarrera.carrera, inscripcion__isnull=False)))
                        ws2.cell(row=(idx + 36), column=3,value=len(solicitudes.filter(revision_director=1, carreradestino=cambiocarrera.carrera, inscripcion__isnull=True)))
                    c2 = BarChart()
                    c2.title = 'CAMBIO DE CARRERA'
                    c2.type = "bar"
                    c2.style = 12
                    data = Reference(ws2, min_col=2, min_row=35, max_row=(len(carreras) + 35))
                    category = Reference(ws2, min_col=1, min_row=36, max_row=(len(carreras) + 35))
                    c2.add_data(data, titles_from_data=True)
                    c2.set_categories(category)
                    ws2.add_chart(c2, 'E35')

                    c3 = BarChart()
                    c3.title = 'CAMBIO DE IES'
                    c3.type = "bar"
                    c3.style = 10
                    c3.shape = 4
                    data = Reference(ws2, min_col=3, min_row=35, max_row=(len(carreras) + 35))
                    category = Reference(ws2, min_col=1, min_row=36, max_row=(len(carreras) + 35))
                    c3.add_data(data, titles_from_data=True)
                    c3.set_categories(category)
                    ws2.add_chart(c3, 'E55')
                    wb.save(response)
                    return response
                except Exception as ex:
                    pass

            elif action == 'estverificacionrequisitos':
                id = int(request.GET['id'])
                filtro = SolicitudCambioCarrera.objects.get(pk=id)
                ESTADOS_PASOS_SOLICITUD_1 = ()
                resp = []
                totaldocumentos = DocumentosSolicitudCambioCarrera.objects.filter(status=True,
                                                                                  solicitud=filtro).count()
                totalaprobados = DocumentosSolicitudCambioCarrera.objects.filter(status=True, solicitud=filtro,
                                                                                 estados=1).count()
                if es_director_adm:
                    if totaldocumentos == totalaprobados and filtro.revision_bienestar == 1 or filtro.puntajeincorrecto :
                        ESTADOS_PASOS_SOLICITUD_1 = (
                            (1, u'APROBADO'),
                            (2, u'RECHAZADO')
                        )
                        resp = [{'id': cr[0], 'text': cr[1]} for cr in ESTADOS_PASOS_SOLICITUD_1]
                    else:
                        resp = [{'id': 2, 'text': 'RECHAZADO'}]
                else:
                    if totaldocumentos == totalaprobados or filtro.puntajeincorrecto:
                        ESTADOS_PASOS_SOLICITUD_1 = (
                            (1, u'APROBADO'),
                            (2, u'RECHAZADO')
                        )
                        resp = [{'id': cr[0], 'text': cr[1]} for cr in ESTADOS_PASOS_SOLICITUD_1]
                    else:
                        resp = [{'id': 2, 'text': 'RECHAZADO'}]
                return HttpResponse(json.dumps({'state': True, 'result': resp}))

            elif action == 'verdocumentos':
                try:
                    data['solicitud'] = solicitud = SolicitudCambioCarrera.objects.get(pk=int(request.GET['id']))
                    ESTADOS_PASOS_SOLICITUD_1 = (
                        (1, u'APROBADO'),
                        (2, u'RECHAZADO')
                    )
                    data['estados'] = ESTADOS_PASOS_SOLICITUD_1
                    ESTADOS_DOCUMENTOS = (
                        (1, u'APROBAR'),
                        (2, u'RECHAZAR'),
                        (3, u'CORREGIR'),
                        (4, u'ELIMINAR')
                    )
                    data['estados_documentos'] = ESTADOS_DOCUMENTOS
                    data['form2'] = SubirEvidenciaCambioCarreraForm()
                    data['roles']=ROLES_CAMBIO_CARRERA
                    data['documentos'] = DocumentosSolicitudCambioCarrera.objects.filter(solicitud=solicitud,
                                                                                         status=True).order_by('documento__fecha_creacion')
                    template = get_template('alu_cambiocarrera/modal/documentossolicitudestudiante.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, "mensaje": ex})

            elif action == 'verdocumentoseliminados':
                try:
                    data['solicitud'] = solicitud = SolicitudCambioCarrera.objects.get(pk=int(request.GET['id']))
                    data['documentos'] = DocumentosSolicitudCambioCarrera.objects.filter(solicitud=solicitud,
                                                                                         status=False).order_by('documento__nombre')
                    template = get_template('alu_cambiocarrera/modal/documentoseliminadossolicitudestudiante.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, "mensaje": ex})

            elif action == 'verproceso':
                try:
                    data['solicitud'] = solicitud = SolicitudCambioCarrera.objects.get(pk=int(request.GET['id']))
                    data['documentos'] = documentos = DocumentosSolicitudCambioCarrera.objects.filter(status=True, solicitud=solicitud).order_by('documento__nombre')
                    data['filtro'] = filtro = solicitud.periodocambiocarrera
                    pasoactual = 1
                    data['paso2'] = paso2 = False if not solicitud.revision_admision == 1 else True
                    data['paso3'] = paso3 = False if not solicitud.revision_bienestar == 1 else True
                    data['paso4'] = paso4 = False if not solicitud.aprobacion_admision == 1 else True
                    data['paso5'] = paso5 = False if not solicitud.revision_decano == 1 else True
                    data['paso6'] = paso6 = False if not solicitud.revision_director == 1 else True
                    if paso2:
                        pasoactual = 2
                    if paso3:
                        pasoactual = 3
                    if paso4:
                        pasoactual = 4
                    if paso5:
                        pasoactual = 5
                    if paso6:
                        pasoactual = 6
                    data['pasoactual'] = pasoactual
                    template = get_template('alu_cambiocarrera/modal/modalproceso.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, "mensaje": ex})

            elif action == 'verseguimiento':
                try:
                    data['filtro'] = SolicitudCambioCarrera.objects.get(pk=int(request.GET['id']))
                    template = get_template('alu_cambiocarrera/modal/verseguimiento.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, "mensaje": ex})

            if action == 'addevidenciavalidada':
                try:
                    data['action']=action
                    data['filtro'] = SolicitudCambioCarrera.objects.get(pk=int(request.GET['id']))
                    template = get_template('alu_cambiocarrera/modal/adddocvalidados.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, "mensaje": ex})

            elif action == 'validardecano':
                try:
                    data['solicitud'] = solicitud = SolicitudCambioCarrera.objects.get(pk=int(request.GET['id']))
                    data['documentos'] = documentos = DocumentosSolicitudCambioCarrera.objects.filter(status=True,
                        solicitud=solicitud).order_by('documento__nombre')
                    data['filtro'] = filtro = solicitud.periodocambiocarrera
                    # data['form2'] = SubirEvidenciaCambioCarreraForm()
                    pasoactual = 1
                    data['paso2'] = paso2 = False if not solicitud.revision_admision == 1 else True
                    data['paso3'] = paso3 = False if not solicitud.revision_bienestar == 1 else True
                    data['paso4'] = paso4 = False if not solicitud.aprobacion_admision == 1 else True
                    data['paso5'] = paso5 = False if not solicitud.revision_decano == 1 else True
                    data['paso6'] = paso6 = False if not solicitud.revision_director == 1 else True
                    if paso2:
                        pasoactual = 2
                    if paso3:
                        pasoactual = 3
                    if paso4:
                        pasoactual = 4
                    if paso5:
                        pasoactual = 5
                    if paso6:
                        pasoactual = 6
                    data['pasoactual'] = pasoactual

                    template = get_template('alu_cambiocarrera/modal/validardecano.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, "mensaje": ex})

            elif action == 'validardirectorcarrera':
                try:
                    data['solicitud'] = solicitud = SolicitudCambioCarrera.objects.get(pk=int(request.GET['id']))
                    data['documentos'] = documentos = DocumentosSolicitudCambioCarrera.objects.filter(status=True,
                        solicitud=solicitud).order_by('documento__nombre')
                    data['filtro'] = filtro = solicitud.periodocambiocarrera
                    data['form2'] = SubirEvidenciaCambioCarreraForm()
                    pasoactual = 1
                    data['paso2'] = paso2 = False if not solicitud.revision_admision == 1 else True
                    data['paso3'] = paso3 = False if not solicitud.revision_bienestar == 1 else True
                    data['paso4'] = paso4 = False if not solicitud.aprobacion_admision == 1 else True
                    data['paso5'] = paso5 = False if not solicitud.revision_decano == 1 else True
                    data['paso6'] = paso6 = False if not solicitud.revision_director == 1 else True
                    if paso2:
                        pasoactual = 2
                    if paso3:
                        pasoactual = 3
                    if paso4:
                        pasoactual = 4
                    if paso5:
                        pasoactual = 5
                    if paso6:
                        pasoactual = 6
                    data['pasoactual'] = pasoactual

                    template = get_template('alu_cambiocarrera/modal/validardirector.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, "mensaje": ex})

            elif action == 'listacarreras':
                try:
                    data['title'] = u'Carreras asignadas'
                    data['periodocabiocarrera'] = apertura = AperturaPeriodoCambioCarrera.objects.get(
                        pk=int(request.GET['id']))
                    data['carreras'] = carreras = apertura.carrerascambiocarrera_set.filter(status=True).order_by(
                        'carrera__nombre')
                    data['cantidad_carreras'] = len(carreras)
                    return render(request, "alu_cambiocarrera/modal/listacarreras.html", data)
                except Exception as e:
                    return JsonResponse({"result": "bad", 'mensaje': str(e)})

            elif action == 'addcarrera':
                try:
                    data['postar'] = apertura = AperturaPeriodoCambioCarrera.objects.get(pk=request.GET['id'])
                    data['yaasignadas'] = detalle = apertura.carrerascambiocarrera_set.filter(status=True).order_by(
                        'carrera__nombre')

                    mallascarrera = Materia.objects.filter(nivel__periodo=periodo).values_list('asignaturamalla__malla__carrera_id', flat=True).distinct()

                    data['carreras'] = Carrera.objects.filter(pk__in=mallascarrera,
                                                              status=True,coordinacion__in=apertura.coordinaciones()).exclude(
                        pk__in=detalle.values_list('carrera_id')).order_by('nombre').distinct()

                    template = get_template('alu_cambiocarrera/modal/addcarreras.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    mensaje = 'Intentelo mas tarde'
                    return JsonResponse({"result": False, "mensaje": mensaje})

            elif action == 'publicarperiodo':
                try:
                    data['postar'] = apertura = AperturaPeriodoCambioCarrera.objects.get(pk=request.GET['id'])
                    aperturaspublico=AperturaPeriodoCambioCarrera.objects.filter(publico=True, status=True)
                    if apertura.publico:
                        apertura.publico = False
                    else:
                        if aperturaspublico:
                            for ap in aperturaspublico:
                                ap.publico = False
                                ap.save(request)
                        apertura.publico = True
                    apertura.save(request)
                except Exception as ex:
                    mensaje = 'Intentelo mas tarde'
                    return JsonResponse({"result": False, "mensaje": mensaje})

            elif action == 'confresponsable':
                data['title'] = u'Configurar Responsables de Revision de Cambio de Carrera'
                search = None
                ids = None
                responsables = ReponsableCambioCarrera.objects.filter(status=True).order_by('-pk')
                if 's' in request.GET:
                    search = request.GET['s'].strip()
                    ss = search.split(' ')
                    if len(ss) == 1:
                        responsables = responsables.filter(Q(persona__nombres__icontains=search) |
                                                                        Q(persona__apellido1__icontains=search) |
                                                                        Q(persona__apellido2__icontains=search) |
                                                                        Q(persona__cedula__icontains=search) |
                                                                        Q(persona__pasaporte__icontains=search) |
                                                                        Q(persona__usuario__username__icontains=search)).distinct()
                    else:
                        responsables = responsables.filter(Q(persona__apellido1__icontains=ss[0]) &
                                                                        Q(persona__apellido2__icontains=ss[1])).distinct()
                paging = MiPaginador(responsables, 10)
                p = 1
                try:
                    paginasesion = 1
                    if 'paginador' in request.session:
                        paginasesion = int(request.session['paginador'])
                    if 'page' in request.GET:
                        p = int(request.GET['page'])
                    else:
                        p = paginasesion
                    try:
                        page = paging.page(p)
                    except:
                        p = 1
                    page = paging.page(p)
                except:
                    page = paging.page(p)
                request.session['paginador'] = p
                data['paging'] = paging
                data['page'] = page
                data['roles']=ROLES_CAMBIO_CARRERA
                # data['idperiodo'] = request.GET['id']
                data['rangospaging'] = paging.rangos_paginado(p)
                data['listarresponsales'] = page.object_list
                data['search'] = search if search else ""
                data['ids'] = ids if ids else ""
                return render(request, "alu_cambiocarrera/confresponsables.html", data)

            elif action == 'addresponsable':
                try:
                    data['action']=action
                    data['roles']=ROLES_CAMBIO_CARRERA
                    template = get_template('alu_cambiocarrera/modal/add&editrolfuncionario.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'editresponsable':
                try:
                    data['action'] = action
                    data['roles'] = ROLES_CAMBIO_CARRERA
                    data['responsable'] = ReponsableCambioCarrera.objects.get(pk=int(request.GET['id']))
                    template = get_template('alu_cambiocarrera/modal/add&editrolfuncionario.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'delresponsable':
                try:
                    data['title'] = u'Eliminar Registro'
                    data['responsable'] = ReponsableCambioCarrera.objects.get(pk=int(encrypt(request.GET['id'])))
                    return render(request, "alu_cambiocarrera/modal/delrolfuncionario.html", data)
                except Exception as ex:
                    pass

            elif action == 'reasignarsolicitud':
                try:
                    data['solicitud']=solicitud=SolicitudCambioCarrera.objects.get(id=request.GET['id'])
                    data['funcionarios'] = personas_admision.exclude(persona_id=solicitud.persona_admision.id)
                    template = get_template('alu_cambiocarrera/modal/reasignarsolicitud.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    mensaje = 'Intentelo mas tarde'
                    return JsonResponse({"result": False, "mensaje": mensaje})

            elif action == 'estadisticas':
                try:
                    data['title'] = 'Estadísticas de Solicitudes'
                    tipografico=1
                    if 'tipografico' in request.GET:
                        tipografico=int(request.GET['tipografico'])
                    if not 'id' in request.GET:
                        raise NameError("Parametro de periodo no encontrado")
                    if not AperturaPeriodoCambioCarrera.objects.filter(pk=int(request.GET['id'])).exists():
                        raise NameError("Periodo no encontrado")
                    data['apertura']=aPeriodo = AperturaPeriodoCambioCarrera.objects.get(pk=int(request.GET['id']))
                    if not SolicitudCambioCarrera.objects.filter(periodocambiocarrera=aPeriodo).exists():
                        raise NameError("No existen solicitudes")
                    data['tipografico']=tipografico
                    # data['solicitudes'] = solicitudes= SolicitudCambioCarrera.objects.filter(periodocambiocarrera=aPeriodo)
                    return render(request, "alu_cambiocarrera/estadistica_solicitudes.html", data)
                except Exception as ex:
                    return HttpResponseRedirect("/alu_cambiocarrera?solicitantes&id={} %s".format(aPeriodo.id) % ex.__str__())

            elif action == 'subiraprobacionadmision':
                try:
                    data['title'] = u'Subir Documento Aprobación Admisión'
                    solicitud = SolicitudCambioCarrera.objects.get(pk=int(request.GET['id']))
                    data['solicitud'] = solicitud

                    template = get_template("alu_cambiocarrera/modal/subiraprobacionadmision.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content, 'title': data['title']})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})


            return HttpResponseRedirect(request.path)
        else:
            try:
                data['title'] = u'Cambios de carrera'
                search = None
                ids = None
                aperturadas = None
                if 'id' in request.GET:
                    ids = request.GET['id']
                    aperturadas = AperturaPeriodoCambioCarrera.objects.filter(pk=ids, status=True).order_by(
                        '-fechaapertura')
                elif 's' in request.GET:
                    search = request.GET['s']
                    s = search.split(" ")
                    if len(s) == 1:
                            aperturadas = AperturaPeriodoCambioCarrera.objects.filter(
                                (Q(motivo__icontains=s[0])) | (Q(mensaje__icontains=s[0])),
                                Q(status=True)).order_by('-fechaapertura')
                    elif len(s) == 2:
                            aperturadas = AperturaPeriodoCambioCarrera.objects.filter(((Q(
                                motivo__icontains=s[0]) & Q(motivo__icontains=s[1])) | (Q(
                                mensaje__icontains=s[0]) & Q(mensaje__icontains=s[1]))), Q(
                                status=True)).order_by('-fechaapertura')
                    elif len(s) == 3:
                            aperturadas = AperturaPeriodoCambioCarrera.objects.filter(((Q(
                                motivo__icontains=s[0]) & Q(motivo__icontains=s[1]) & Q(
                                motivo__icontains=s[2])) | (Q(mensaje__icontains=s[0]) & Q(
                                mensaje__icontains=s[1]) & Q(mensaje__icontains=s[2]))), Q(
                                status=True)).order_by('-fechaapertura')
                else:
                    aperturadas = AperturaPeriodoCambioCarrera.objects.filter(status=True).order_by(
                        '-fechaapertura')
                paging = MiPaginador(aperturadas, 10)
                p = 1
                try:
                    paginasesion = 1
                    if 'paginador' in request.session:
                        paginasesion = int(request.session['paginador'])
                    if 'page' in request.GET:
                        p = int(request.GET['page'])
                    else:
                        p = paginasesion
                    try:
                        page = paging.page(p)
                    except:
                        p = 1
                    page = paging.page(p)
                except:
                    page = paging.page(p)
                request.session['paginador'] = p
                data['totalasignado']=SolicitudCambioCarrera.objects.filter(persona_admision=persona.id).count()
                data['paging'] = paging
                data['page'] = page
                data['funcionarios']=ReponsableCambioCarrera.objects.filter(status=True, rol=0).exists()
                data['rangospaging'] = paging.rangos_paginado(p)
                data['listasapertura'] = page.object_list
                data['search'] = search if search else ""
                data['ids'] = ids if ids else ""
                actualizaperiodos()
                return render(request, "alu_cambiocarrera/viewperiodocambiocarrera.html", data)
            except Exception as ex:
                response = JsonResponse({'result': False, 'mensaje': 'ERROR EN LA TRANSACCION'})


def actualizaperiodos():
    if AperturaPeriodoCambioCarrera.objects.filter(status=True, publico=True).exists():
        if AperturaPeriodoCambioCarrera.objects.filter(status=True, publico=True).count() > 1:
            for per in AperturaPeriodoCambioCarrera.objects.filter(status=True, publico=True):
                if not per.esta_en_fechas():
                    per.publico=False
                    per.save()
        pass

