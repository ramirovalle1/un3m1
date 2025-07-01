# -*- coding: UTF-8 -*-
import calendar
import json

import requests
from django.db.models import Count
from django.forms import model_to_dict
from openpyxl import workbook as openxl
from openpyxl.chart import ScatterChart, Reference, Series, PieChart, BarChart
from django.template.loader import get_template
from openpyxl.styles import Font as openxlFont
from openpyxl.styles.alignment import Alignment as alin
from sga.funcionesxhtml2pdf import conviert_html_to_2pdf

from xlwt import *

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.shortcuts import render, redirect
from decorators import secure_module
from settings import EMAIL_DOMAIN
from sga.commonviews import adduserdata
from sga.templatetags.sga_extras import encrypt
from sga.funciones import MiPaginador, log, notificacion, generar_nombre
from utils.filtros_genericos import filtro_persona_select, consultarPersona
from .forms import FinalizaCitaForm, ValidarCitaForm, GestionServicioCitaForm, CitaEmergenteForm,RefuerzoAcademicoForm
from .models import *
from django.db.models import Sum, Q, F, FloatField
from cita.forms import PersonaForm, InformePsicologicoForm
from sagest.forms import FamiliarForm
from sga.models import Persona, PersonaDatosFamiliares
from django import template, forms
from sagest.funciones import encrypt_id
from datetime import datetime

@login_required(redirect_field_name='ret', login_url='/loginsga')
@secure_module
@transaction.atomic()
def view(request):
    data = {}
    adduserdata(request, data)
    hoy = datetime.now().date()
    usuario = request.user

    data['persona']= persona = request.session['persona']
    periodo = request.session['periodo']
    perfilprincpal = request.session['perfilprincipal']
    data['es_responsable'] = es_responsable = persona.responsableserviciocita_set.filter(status=True).exists()
    # if not es_responsable and not request.user.has_perm('sga.puede_revisar_total_citas'):
    if (not es_responsable and not request.user.has_perm(
            'sga.puede_revisar_total_citas')):
        messages.error(request, 'Usted no tiene acceso a este modulo')
        return redirect('/')
    if request.method == 'POST':
        res_json = []
        action = request.POST['action']
        if action == 'finalizar':
            with transaction.atomic():
                try:
                    id = int(encrypt(request.POST['id']))
                    ids = int(encrypt(request.POST['ids']))
                    if id != 0:
                        instance = PersonaCitaAgendada.objects.get(pk=id)
                        cita=instance
                    elif ids != 0:
                        instance = SubCitaAgendada.objects.get(pk=ids)
                        cita = instance.citaprincipal
                    form = FinalizaCitaForm(request.POST)
                    if form.is_valid():
                        instance.estado = int(form.cleaned_data['estado'])
                        instance.observacion = form.cleaned_data['observacion']
                        instance.asistio = form.cleaned_data['asistio']

                        if instance.estado == 5 and cita.escitaemergente and id == 0:
                            instance.termina =datetime.now().time()
                        instance.save(request)

                        if id != 0 and int(instance.estado) == 5:
                            for subcita in cita.subcitas():
                                if subcita.estado == 1 or subcita.estado == 6:
                                    subcita.estado = 5
                                    subcita.save(request)
                        if ids != 0 and form.cleaned_data['finalizar']:
                            cita.estado = 5
                            cita.save(request)
                            for subcita in cita.subcitas():
                                if subcita.estado == 1 or subcita.estado == 6:
                                    subcita.estado = 5
                                    subcita.save(request)
                        servicio=cita.servicio.serviciocita.nombre.lower()

                        if instance.estado == 2 or instance.estado == 6:
                            titulo = u"Culminación de cita agendada en {} - ({})".format(servicio, cita.get_estado_display())
                            if instance.estado == 2:
                                mensaje = 'Su cita agendada en {} fue {}.'.format(servicio, cita.get_estado_display())
                            else:
                                mensaje = 'Su cita agendada en {} fue puesto {}.'.format(servicio, cita.get_estado_display().lower())
                            notificacion(titulo, mensaje, cita.persona, None,
                                         f'/alu_agendamientocitas?action=miscitas&search={cita.codigo}',
                                         instance.pk, 1, 'sga', PersonaCitaAgendada, request)
                            lista_email = cita.persona.lista_emails()
                            # lista_email = ['jguachuns@unemi.edu.ec', ]
                            datos_email = {'sistema': request.session['nombresistema'],
                                           'fecha': datetime.now().date(),
                                           'hora': datetime.now().time(),
                                           'cita': instance,
                                           'persona': cita.persona,
                                           'mensaje': mensaje}
                            template = "emails/notificacion_agendamientocitas.html"
                            send_html_mail(titulo, template, datos_email, lista_email, [], [],
                                           cuenta=CUENTAS_CORREOS[0][1])
                        #Guardo información para historial.
                        historial = HistorialSolicitudCita(cita=cita, estado_solicitud=instance.estado, observacion=instance.observacion)
                        historial.save(request)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                             "mensaje": "Error en el formulario"})
                    log(u'Cambió estado a cita: %s' % instance, request, "finalizar")
                    return JsonResponse({"result": False}, safe=False)
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": True, "mensaje": "Intentelo más tarde: {}".format(str(ex))}, safe=False)

        if action == 'delcita':
            with transaction.atomic():
                try:
                    instance = PersonaCitaAgendada.objects.get(pk=int(encrypt(request.POST['id'])))
                    instance.status = False
                    instance.save(request)

                    for cita in instance.subcitas():
                        cita.status=False
                        cita.save(request)
                    historial = HistorialSolicitudCita(cita=instance, observacion='Eliminación de cita')
                    historial.save(request)
                    log(u'Elimino cita: %s' % instance, request, "delcita")
                    res_json = {"error": False}
                except Exception as ex:
                    res_json = {'error': True, "message": "Error: {}".format(ex)}
                return JsonResponse(res_json, safe=False)

        if action == 'mostrarservicio':
            with transaction.atomic():
                try:
                    registro = GestionServicioCita.objects.get(pk=request.POST['id'])
                    registro.mostrar = eval(request.POST['val'].capitalize())
                    registro.save(request)
                    log(u'Mostrar servicio: %s (%s)' % (registro, registro.mostrar), request, "edit")
                    return JsonResponse({"result": True, 'mensaje': 'Cambios guardados'})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": False})

        if action == 'validar':
            with transaction.atomic():
                try:
                    instance = DocumentosSolicitudServicio.objects.get(pk=int(request.POST['id']))
                    form = ValidarCitaForm(request.POST)
                    if form.is_valid():
                        instance.estados = int(form.cleaned_data['estado'])
                        instance.observacion = form.cleaned_data['observacion']
                        instance.save(request)

                        cita = instance.cita
                        servicio=cita.servicio.serviciocita.nombre.lower()
                        documentos = DocumentosSolicitudServicio.objects.filter(cita=cita).exclude(id=instance.id)
                        mensaje =''
                        if instance.estados == 1 and cita.doc_validacion() == 1 and not cita.subcitas_exits() and not cita.estado == 6:
                            cita.estado=1
                            cita.save(request)
                            mensaje='Los documentos subidos en la cita agendada de {} fueron aprobados y su cita fue agendada'.format(cita.servicio.serviciocita.nombre.capitalize())
                        elif instance.estados == 1 and cita.doc_validacion() == 1 and cita.subcitas_exits():
                            cita.estado = 6
                            cita.save(request)
                            mensaje = 'Los documentos subidos en la cita agendada de {} fueron aprobados'.format(cita.servicio.serviciocita.nombre.capitalize())
                        elif cita.doc_validacion() == 0 and cita.estado != 0:
                            cita.estado = 0
                            cita.save(request)
                        elif instance.estados == 2:
                            cita.estado=3
                            cita.save(request)
                            mensaje = 'Su cita agendada en {} tiene documentos por corregir.'.format(servicio)
                        if instance.estados == 3:
                            mensaje = 'Su cita agendada en {} tiene documentos rechazados.'.format(servicio)

                        if instance.estados == 1 and cita.doc_validacion() == 1 or instance.estados != 1:
                            titulo = u"Validación de documentos en cita agendada de {} - ({})".format(servicio, instance.get_estados_display())
                            notificacion(titulo,
                                         mensaje,cita.persona, None, f'/alu_agendamientocitas?action=miscitas&search={cita.codigo}',
                                         instance.pk, 1, 'sga', DocumentosSolicitudServicio, request)
                            lista_email = cita.persona.lista_emails()
                            # lista_email = ['jguachuns@unemi.edu.ec', ]
                            datos_email = {'sistema': request.session['nombresistema'],
                                           'fecha': datetime.now().date(),
                                           'hora': datetime.now().time(),
                                           'documento': instance,
                                           'persona': cita.persona,
                                           'mensaje': mensaje}
                            template = "emails/notificacion_agendamientocitas.html"
                            send_html_mail(titulo, template, datos_email, lista_email, [], [],
                                           cuenta=CUENTAS_CORREOS[0][1])
                        #Guardo informacion de historial de agendamiento cita
                        historial=HistorialSolicitudCita(cita=cita, documento=instance, estado_documento=instance.estados, estado_solicitud=cita.estado, observacion=instance.observacion)
                        historial.save(request)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                             "mensaje": "Error en el formulario"})
                    diccionario={'id':instance.id,'observacion':instance.observacion, 'estado':instance.get_estados_display(),'idestado':instance.estados, 'color':instance.color_estado()}
                    log(u'Valido documento de cita: %s' % instance, request, "edit")
                    return JsonResponse({'result': True,'data_return':True, 'mensaje': u'Guardado con exito', 'data':diccionario}, safe=False)
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": True, "mensaje": "Error {}".format(str(ex))}, safe=False)

        if action =='addgestion':
            with transaction.atomic():
                try:
                    id=int(encrypt(request.POST['id']))
                    ids=int(encrypt(request.POST['ids']))
                    form = GestionServicioCitaForm(request.POST, request.FILES)
                    if form.is_valid():
                        if id != 0:
                            instance = PersonaCitaAgendada.objects.get(pk=id)
                            gestion=GestionServicioCita(cita=instance, observacion=form.cleaned_data['observacion'], mostrar=form.cleaned_data['mostrar'])
                        elif ids != 0:
                            instance = SubCitaAgendada.objects.get(pk=ids)
                            gestion=GestionServicioCita(subcita=instance, observacion=form.cleaned_data['observacion'],mostrar=form.cleaned_data['mostrar'])

                        gestion.save(request)
                        if 'archivo' in request.FILES:
                            newfile = request.FILES['archivo']
                            extension = newfile._name.split('.')
                            tam = len(extension)
                            exte = extension[tam - 1]
                            if newfile.size > 4194304:
                                transaction.set_rollback(True)
                                return JsonResponse(
                                    {"result": True, "mensaje": u"Error, el tamaño del archivo es mayor a 2 Mb."})
                            if not exte.lower() in ['pdf']:
                                transaction.set_rollback(True)
                                return JsonResponse({"result": True, "mensaje": u"Error, solo archivos .pdf"})
                            newfile._name = generar_nombre(
                                "Consentimiento_informado_{}_{}".format(gestion.pk, random.randint(1, 100000).__str__()),
                                newfile._name)
                            gestion.archivo = newfile
                            gestion.save(request)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                             "mensaje": "Error en el formulario"})
                    log(u'Adiciono observacion a cita: %s' % instance, request, "add")
                    return JsonResponse({"result": False}, safe=False)
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": True, "mensaje": "Intentelo más tarde: {}".format(str(ex))}, safe=False)

        if action =='editgestion':
            with transaction.atomic():
                try:
                    form = GestionServicioCitaForm(request.POST, request.FILES)
                    if form.is_valid():
                        gestion=GestionServicioCita.objects.get(id=int(encrypt(request.POST['ide'])))
                        gestion.observacion=form.cleaned_data['observacion']
                        gestion.save(request)
                        if 'archivo' in request.FILES:
                            newfile = request.FILES['archivo']
                            extension = newfile._name.split('.')
                            tam = len(extension)
                            exte = extension[tam - 1]
                            if newfile.size > 4194304:
                                transaction.set_rollback(True)
                                return JsonResponse(
                                    {"result": True, "mensaje": u"Error, el tamaño del archivo es mayor a 2 Mb."})
                            if not exte.lower() in ['pdf']:
                                transaction.set_rollback(True)
                                return JsonResponse({"result": True, "mensaje": u"Error, solo archivos .pdf"})
                            newfile._name = generar_nombre("Respaldo_observacion_{}_{}".format(gestion.pk, random.randint(1, 100000).__str__()),newfile._name)
                            gestion.archivo = newfile
                            gestion.save(request)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                             "mensaje": "Error en el formulario"})
                    log(u'Edito observacion de gestion de cita: %s' % gestion, request, "edit")
                    return JsonResponse({"result": False}, safe=False)
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": True, "mensaje": "Intentelo más tarde: {}".format(str(ex))}, safe=False)

        if action == 'delgestion':
            with transaction.atomic():
                try:
                    gestion = GestionServicioCita.objects.get(id=int(encrypt(request.POST['id'])))
                    gestion.status=False
                    gestion.save(request)
                    log(u'Elimino Gestion, observacion de reserva: %s' % gestion, request, "delobservacion")
                    res_json = {"error": False}
                except Exception as ex:
                    res_json = {'error': True, "message": "Error: {}".format(ex)}
                return JsonResponse(res_json, safe=False)

        if action == 'addcita':
            with transaction.atomic():
                try:
                    fecha = datetime.strptime(request.POST['fecha'], '%Y-%m-%d').date()
                    idservicio=int(encrypt(request.POST['idservicio']))
                    idcita=int(encrypt(request.POST['idcita']))
                    solicitar=request.POST.getlist('solicitar[]')
                    if not request.POST['horario']:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Seleccione un horario a reservar."})
                    horario = HorarioServicioCita.objects.get(id=request.POST['horario'])
                    servicio = ServicioConfigurado.objects.get(id=idservicio)
                    tipo_atencion=horario.tipo_atencion
                    if tipo_atencion == 0:
                        tipo_atencion=request.POST['tipo_atencion']
                    citaprincipal=PersonaCitaAgendada.objects.get(id=idcita)
                    # turno = horario.generar_turno(str(persona))
                    derivacion = not citaprincipal.persona_responsable == horario.responsableservicio.responsable
                    cita = SubCitaAgendada(citaprincipal=citaprincipal,
                                           servicio=servicio,
                                           horario=horario,
                                           estado=1,
                                           fechacita=fecha,
                                           persona_responsable=horario.responsableservicio.responsable,
                                           es_derivacion=derivacion,
                                           tipo_atencion=tipo_atencion)
                    cita.save(request)
                    requisitos = RequisitoServicioCita.objects.filter(status=True, id__in=solicitar)
                    for rs in requisitos:
                        documento = DocumentosSolicitudServicio.objects.filter(status=True, cita=citaprincipal, requisito__requisito=rs.requisito).first()
                        if documento:
                            documento.obligatorio = True
                            documento.save(request)
                        else:
                            docrequerido = DocumentosSolicitudServicio(cita=citaprincipal, obligatorio=True, requisito=rs,
                                                                       estados=0)
                            docrequerido.save(request)
                    servicio=cita.citaprincipal.servicio.serviciocita.nombre.lower()
                    titulo = u"Agendamiento de cita en {} ({} | {})".format(servicio, cita.fechacita,cita.horario.turno.nombre_horario())
                    mensaje = 'Se le agendo una nueva cita para su atención.'
                    lista_email = citaprincipal.persona.lista_emails()
                    # lista_email = ['jguachuns@unemi.edu.ec', ]
                    datos_email = {'sistema': request.session['nombresistema'],
                                   'fecha': datetime.now().date(),
                                   'hora': datetime.now().time(),
                                   'subcita': cita,
                                   'persona': citaprincipal.persona,
                                   'mensaje': mensaje,
                                   'requisitos_sol':requisitos}
                    template = "emails/notificacion_agendamientocitas.html"
                    send_html_mail(titulo, template, datos_email, lista_email, [], [], cuenta=CUENTAS_CORREOS[0][1])
                    # Guardo informacion de historial de agendamiento cita
                    historial = HistorialSolicitudCita(cita=citaprincipal, estado_solicitud=cita.estado, observacion='Creación de sub cita')
                    historial.save(request)
                    messages.success(request,'Se agendo la cita correctamente.')
                    log(u'{} : Inicio Agendamiento de sub Cita  {}'.format(persona,horario.responsableservicio.servicio.__str__()),request,"add")
                    url_ = '{}?action=gestionarsubcita&id={}'.format(request.path,encrypt(cita.id))
                    return JsonResponse({"result": False, "to": url_})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "mensaje": 'Error: {}'.format(ex)}, safe=False)

        if action == 'solicitar':
            with transaction.atomic():
                try:
                    instance = DocumentosSolicitudServicio.objects.get(pk=int(request.POST['id']))
                    instance.obligatorio=True
                    instance.save(request)

                    cita = instance.cita
                    responsable = cita.persona_responsable.nombre_normal_minus()
                    requisito_ = instance.requisito.requisito.nombre.capitalize()
                    titulo = u"Subir documento de cita agendada ({})".format('Subir archivo')
                    mensaje = 'Documento ({}) fue solicitado por el responsable {} para ser subido a la plataforma, por favor revisar su cita y subir los requisitos solicitados.'.format(requisito_, responsable)
                    notificacion(titulo, mensaje, cita.persona, None,
                                 f'/alu_agendamientocitas?action=miscitas&search={cita.codigo}',
                                 instance.pk, 1, 'sga', DocumentosSolicitudServicio, request)
                    titulo = titulo
                    lista_email = cita.persona.lista_emails()
                    # lista_email = ['jguachuns@unemi.edu.ec', ]
                    datos_email = {'sistema': request.session['nombresistema'],
                                   'fecha': datetime.now().date(),
                                   'hora': datetime.now().time(),
                                   'documento': instance,
                                   'persona': cita.persona,
                                   'mensaje': mensaje,}
                    template = "emails/notificacion_agendamientocitas.html"
                    send_html_mail(titulo, template, datos_email, lista_email, [], [], cuenta=CUENTAS_CORREOS[0][1])

                    diccionario = {'id': instance.id, 'observacion': instance.observacion,
                                   'estado': 'Sin archivo', 'idestado': instance.estados,
                                   'color': instance.color_estado(), 'obl_':instance.obligatorio}
                    log(u'Solicito documento de cita: %s' % instance, request, "edit")
                    return JsonResponse({'result': True, 'data_return': True, 'mensaje': u'Guardado con exito', 'data': diccionario},safe=False)
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": True, "mensaje": "Intentelo más tarde: {}".format(str(ex))}, safe=False)

        if action == 'admaddcita':
            with transaction.atomic():
                try:
                    if not request.POST['horario']:
                        raise NameError("Seleccione un horario para agendar.")
                    if int(encrypt(request.POST['persona_ag'])) == 0:
                        raise NameError("Seleccione una persona.")
                    if int(encrypt(request.POST['idperfil'])) == 0:
                        raise NameError("Seleccione un perfil de usuario para agendar la cita.")
                    solicitar = request.POST.getlist('solicitar[]')
                    idpersona = int(encrypt(request.POST['persona_ag']))
                    idperfil = int(encrypt(request.POST['idperfil']))
                    fecha = datetime.strptime(request.POST['fecha'], '%Y-%m-%d').date()
                    idservicio = int(encrypt(request.POST['idservicio']))
                    horario = HorarioServicioCita.objects.get(id=request.POST['horario'])
                    tipo_atencion = horario.tipo_atencion
                    if horario.tipo_atencion == 0:
                        tipo_atencion = request.POST['tipo_atencion']
                    # if horario.requisitos_archivo().filter(opcional=False):
                    #     estado = 0
                    persona_ag = Persona.objects.get(id=idpersona)
                    turno = horario.generar_turno(persona_ag)
                    familiar = None
                    paraFamiliar = False
                    if 'esFamiliar' in request.POST:
                        paraFamiliar = True
                        idFamiliar = int(request.POST['id_familiar'])
                        familiar = PersonaDatosFamiliares.objects.get(id=idFamiliar)

                    cita = PersonaCitaAgendada(persona=persona_ag,
                                               persona_responsable=horario.responsableservicio.responsable,
                                               servicio_id=idservicio,
                                               horario=horario,
                                               estado=1,
                                               perfil_id=idperfil,
                                               codigo=turno,
                                               fechacita=fecha,
                                               espersonal=paraFamiliar,
                                               familiar=familiar,
                                               tipo_atencion=tipo_atencion)
                    cita.save(request)
                    # Guardo informacion de historial de agendamiento cita
                    historial = HistorialSolicitudCita(cita=cita, estado_solicitud=cita.estado,
                                                       observacion='Creación de cita')
                    historial.save(request)

                    # Algoritmo de guardado de documentos requeridos
                    requisitos_solicitados = horario.requisitos_archivo()
                    for rs in requisitos_solicitados:
                        obligatorio = False
                        if solicitar and str(rs.id) in solicitar:
                            obligatorio = True
                        docrequerido = DocumentosSolicitudServicio(cita=cita, obligatorio=obligatorio, requisito=rs,
                                                                   estados=0)
                        docrequerido.save(request)

                    requisitos_sol = RequisitoServicioCita.objects.filter(id__in=solicitar)
                    servicio = cita.servicio.serviciocita.nombre.lower()
                    titulo = u"Agendamiento de cita en {} ({} | {})".format(servicio, cita.fechacita,
                                                                            cita.horario.turno.nombre_horario())
                    mensaje = f'Se le agendo una nueva cita en {servicio}'

                    notificacion(mensaje, titulo, cita.persona, None,
                                 f'/alu_agendamientocitas?action=miscitas&search={cita.codigo}',
                                 cita.pk, 1, 'sga', PersonaCitaAgendada, request)
                    lista_email = cita.persona.lista_emails()
                    # lista_email = ['jguachuns@unemi.edu.ec', ]
                    datos_email = {'sistema': request.session['nombresistema'],
                                   'fecha': datetime.now().date(),
                                   'hora': datetime.now().time(),
                                   'admcita': cita,
                                   'persona': cita.persona,
                                   'mensaje': mensaje,
                                   'requisitos_sol': requisitos_sol}
                    template = "emails/notificacion_agendamientocitas.html"
                    send_html_mail(titulo, template, datos_email, lista_email, [], [], cuenta=CUENTAS_CORREOS[0][1])

                    lista_email = cita.persona_responsable.lista_emails()
                    # lista_email = ['jguachuns@unemi.edu.ec', ]
                    mensaje = f'Una cita fue agendada por {cita.persona.nombre_completo_minus()} para su atención.'
                    datos_email = {'sistema': request.session['nombresistema'],
                                   'fecha': datetime.now().date(),
                                   'hora': datetime.now().time(),
                                   'subcita': cita,
                                   'persona': cita.persona_responsable,
                                   'mensaje': mensaje}
                    template = "emails/notificacion_admin_agendamientocitas.html"
                    send_html_mail(titulo, template, datos_email, lista_email, [], [], cuenta=CUENTAS_CORREOS[0][1])
                    # Guardo informacion de historial de agendamiento cita
                    historial = HistorialSolicitudCita(cita=cita, estado_solicitud=cita.estado,
                                                       observacion='Creación de sub cita')
                    historial.save(request)
                    messages.success(request, 'Se agendo la cita correctamente.')
                    log(u'{} : Inicio Agendamiento de sub Cita  {}'.format(persona,horario.responsableservicio.servicio.__str__()),
                        request, "add")
                    url_ = '{}?action=admagendar'.format(request.path)
                    return JsonResponse({"result": False, "to": url_})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "mensaje": 'Error: {}'.format(ex)}, safe=False)

        if action == 'addcitaemergente':
            with transaction.atomic():
                try:
                    form = CitaEmergenteForm(request.POST)
                    if not form.is_valid():
                        transaction.set_rollback(True)
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                             "mensaje": "Error en el formulario"})
                    servicioC = ServicioConfigurado.objects.get(id = request.POST['servicio'])
                    # persona_ag = Persona.objects.get(id=form.cleaned_data['persona'])
                    turno = servicioC.generar_turno(form.cleaned_data['persona'])
                    esFamiliar = form.cleaned_data['mostrar']
                    familiar = None
                    if esFamiliar:
                        familiar = form.cleaned_data['personafamiliar']
                    cita = PersonaCitaAgendada(persona=form.cleaned_data['persona'],
                                               perfil=form.cleaned_data['perfil'],
                                               servicio=form.cleaned_data['servicio'],
                                               fechacita=hoy,
                                               persona_responsable=form.cleaned_data['persona_responsable'],
                                               tipo_atencion = form.cleaned_data['tipo_atencion'],
                                               comienza = form.cleaned_data['comienza'],
                                               estado = 1,
                                               codigo = turno,
                                               espersonal = esFamiliar,
                                               familiar = familiar,
                                               escitaemergente = True)

                    cita.save(request)
                    # Guardo informacion de historial de agendamiento cita
                    historial = HistorialSolicitudCita(cita=cita, estado_solicitud=cita.estado, observacion='Creación de cita emergente')
                    historial.save(request)


                    return JsonResponse({"result": False,})

                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "mensaje": 'Error: {}'.format(ex)}, safe=False)
                return JsonResponse({"result": "bad", "mensaje": u"Solicitud Incorrecta."})

        # Registrar persona externa
        elif action == 'addpersona':
            try:
                f = PersonaForm(request.POST)
                if not f.is_valid():
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "mensaje": "Error en el formulario"})
                identificacion = f.cleaned_data['identificacion'].upper().strip()
                pers = consultarPersona(identificacion)
                if not pers:
                    pers = Persona(cedula=identificacion,
                                   nombres=f.cleaned_data['nombre'],
                                   apellido1=f.cleaned_data['apellido1'],
                                   apellido2=f.cleaned_data['apellido2'],
                                   nacimiento=f.cleaned_data['nacimiento'],
                                   telefono=f.cleaned_data['telefono'],
                                   sexo=f.cleaned_data['sexo'],
                                   email=f.cleaned_data['email'],
                                   telefono_conv=f.cleaned_data['telefono_conv'],
                                   )
                    pers.save(request)
                    log(u'Adiciono persona: %s' % persona, request, "add")
                else:
                    if not len(pers.mis_perfilesusuarios()) == 1 or not pers.tiene_usuario_externo():
                        raise NameError('Identificación que intenta registrar ya se encuentra registrado')
                    pers.nombres = f.cleaned_data['nombre']
                    pers.apellido1 = f.cleaned_data['apellido1']
                    pers.apellido2 = f.cleaned_data['apellido2']
                    pers.nacimiento = f.cleaned_data['nacimiento']
                    pers.telefono = f.cleaned_data['telefono']
                    pers.sexo = f.cleaned_data['sexo']
                    pers.telefono_conv = f.cleaned_data['telefono_conv']
                    pers.email = f.cleaned_data['email']
                    pers.save(request)
                    log(u'Edito persona de usuario: %s' % pers, request, "edit")
                if not pers.tiene_perfil():
                    externo = Externo.objects.filter(persona=pers, status=True).first()
                    if not externo:
                        externo = Externo(persona=pers)
                        externo.save(request)
                        log(u'Adiciono externo: %s' % externo, request, "add")
                    perfil = PerfilUsuario(persona=pers, externo=externo)
                    perfil.save(request)
                    log(u'Adiciono perfil de usuario externo: %s' % perfil, request, "add")
                return JsonResponse({'result': False, 'mensaje': 'Guardado con éxito'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'{ex}'})

        elif action == 'delhistorialclinico':
            with transaction.atomic():
                try:
                    historial = InformePsicologico.objects.get(id=int(request.POST['id']))
                    historial.status=False
                    historial.save(request)
                    log(u'Elimino Historial: %s' % historial, request, "delhistorialclinico")
                    res_json = {"error": False}
                except Exception as ex:
                    res_json = {'error': True, "message": "Error: {}".format(ex)}
                return JsonResponse(res_json, safe=False)

        elif action == 'delfamiliar':
            try:
                # persona = request.session['persona']
                familiar = PersonaDatosFamiliares.objects.get(pk=encrypt_id(request.POST['id']))
                familiar.status = False
                familiar.save(request)
                log(u'Elimino familiar: %s' % persona, request, "del")
                res_js = {'error': False}
            except Exception as ex:
                transaction.set_rollback(True)
                err = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
                msg_err = f"Ocurrio un error: {ex.__str__()}. {err}"
                res_js = {'error': True, 'mensaje': msg_err}
            return JsonResponse(res_js)

        # Registrar historial
        if action == 'addinfomepsicologia':
            with (transaction.atomic()):
                try:
                    repuestaproceso = json.loads(request.POST['lista_items1'])
                    repuestadetalle = repuestaproceso[-1]  # obtenemos solo el último elemento
                    repuestaproceso = repuestaproceso[:-1]  # obviamos el último elemento
                    form = InformePsicologicoForm(request.POST)

                    if form.is_valid():  # and form.validador():

                        idCab = int(request.POST['ide'])
                        tipoinforme = int(request.POST['tipoinforme'])
                        niveltitulacion = int(request.POST['niveltitulacion'])
                        cita = PersonaCitaAgendada.objects.get(id=int(request.POST['persona']))
                        motivocita = MotivoCita.objects.get(id=int(request.POST['motivoconsulta']))
                        secuencial = InformePsicologico.obtener_ultimo_secuencial_con_ceros(
                            cita.servicio.serviciocita.id)
                        # nivel = NivelTitulacion.objects.get(pk=int(request.POST['niveltitulacion']))

                        # Guardar el valor de institucioneducativa como texto
                        institucion = request.POST['institucioneducativa'] if request.POST[
                            'institucioneducativa'] else None

                        if idCab == 0:  # agregar nuevo
                            instancia = InformePsicologico(
                                codigo=secuencial,
                                personacita=cita,
                                motivoconsulta=motivocita,
                                descripcionmotivoconsulta=form.cleaned_data['descripcionmotivoconsulta'],
                                institucioneducativa=institucion,  # ahora como texto
                                niveltitulacion=niveltitulacion,
                                tipoinforme=tipoinforme,
                                grado=form.cleaned_data['grado']
                            )
                            instancia.save(request)
                        else:  # actualizar cabecera existente
                            instancia = InformePsicologico.objects.get(id=idCab)

                            detalles_procesos = DetalleHistorialPsicologico.objects.filter(informe=instancia)
                            detalles_procesos.update(status=False)

                            detalles_informe = DetalleInformePsicologico.objects.filter(informe=instancia)
                            detalles_informe.update(status=False)

                            instancia.niveltitulacion = niveltitulacion
                            instancia.institucioneducativa = institucion  # ahora como texto
                            instancia.descripcionmotivoconsulta = form.cleaned_data['descripcionmotivoconsulta']
                            instancia.motivoconsulta = motivocita
                            instancia.grado = int(request.POST['grado'])
                            instancia.save(request)

                        # Procesar los datos de repuestaproceso y guardar en DetalleHistorialCitas
                        for respuesta in repuestaproceso:
                            idproceso = int(respuesta['id_proceso'])
                            observacion = respuesta.get('observacion', '')
                            marcado = respuesta['marcado']

                            # Crear una instancia de DetalleHistorialPsicologico
                            detalle_historial = DetalleHistorialPsicologico(
                                informe=instancia,
                                proceso_id=idproceso,
                                observacion=observacion,
                                marcada=marcado
                            )
                            detalle_historial.save()

                        # Procesar detalles del informe
                        for respuesta in repuestadetalle:
                            estruc = EstructuraInforme.objects.get(id=int(respuesta['id']))
                            detalle = respuesta['respuesta']
                            detalle_estructura = DetalleInformePsicologico(
                                descripcion=detalle,
                                estructura=estruc,
                                informe=instancia
                            )
                            detalle_estructura.save()

                        # Manejo de archivo
                        if 'archivo' in request.FILES:
                            newfile = request.FILES['archivo']
                            extension = newfile._name.split('.')
                            exte = extension[-1].lower()
                            if newfile.size > 4194304:
                                transaction.set_rollback(True)
                                return JsonResponse(
                                    {"result": True, "mensaje": u"Error, el tamaño del archivo es mayor a 4 Mb."})
                            if exte not in ['pdf']:
                                transaction.set_rollback(True)
                                return JsonResponse({"result": True, "mensaje": u"Error, solo archivos .pdf"})
                            newfile._name = generar_nombre(
                                "Respaldo_observacion_{}_{}".format(instancia.pk, random.randint(1, 100000)),
                                newfile._name)
                            instancia.archivo = newfile
                            instancia.save(request)

                        log(u'Adicionó Historial de Citas: %s' % instancia, request, "add")
                        return JsonResponse({'result': False, 'mensaje': 'Guardado con éxito'})

                    else:
                        transaction.set_rollback(True)
                        return JsonResponse(
                            {'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                             "mensaje": "Error en el formulario"})

                except Exception as ex:
                    print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "mensaje": 'Error: {}'.format(ex)}, safe=False)
        # if action == 'addinfomepsicologia':
        #     with (transaction.atomic()):
        #         try:
        #             repuestaproceso = json.loads(request.POST['lista_items1'])
        #             repuestadetalle = repuestaproceso[-1]  # obtenemos solo el ultimo elemento
        #             repuestaproceso = repuestaproceso[:-1] #obviamos el ultimo elemento
        #             form = InformePsicologicoForm(request.POST)
        #
        #             if form.is_valid():  # and form.validador():
        #
        #                 idCab = int(request.POST['ide'])
        #                 tipoinforme = int(request.POST['tipoinforme'])
        #                 cita = PersonaCitaAgendada.objects.get(id=int(request.POST['persona']))
        #                 # fechacita = cita.fechacita
        #                 motivocita = MotivoCita.objects.get(id=int(request.POST['motivoconsulta']))
        #                 secuencial = InformePsicologico.obtener_ultimo_secuencial_con_ceros(cita.servicio.serviciocita.id)
        #                 nivel = NivelTitulacion.objects.get(pk=int(request.POST['niveltitulacion']))
        #                 if request.POST['institucioneducativa'] == '':
        #                     institucion = None
        #                 else:
        #                     institucion = InstitucionesColegio.objects.get(pk=int(request.POST['institucioneducativa']))
        #                 # tipoinforme = cita.servicio.serviciocita
        #
        #                 if idCab == 0: #agrega nuevo
        #                     instancia = InformePsicologico(codigo=secuencial,
        #                                                       personacita=cita,
        #                                                       motivoconsulta=motivocita,
        #                                                       descripcionmotivoconsulta=form.cleaned_data['descripcionmotivoconsulta'],
        #                                                       institucioneducativa=institucion,
        #                                                       niveltitulacion=nivel,
        #                                                       tipoinforme = tipoinforme,
        #                                                       grado=form.cleaned_data['grado'])
        #                     instancia.save(request)
        #                 else: #ya existe asi que actualiza cabecera
        #                     instancia = InformePsicologico.objects.get(id=idCab)
        #
        #                     detalles_procesos = DetalleHistorialPsicologico.objects.filter(informe=instancia)
        #                     detalles_procesos.update(status=False)
        #
        #                     detalles_informe = DetalleInformePsicologico.objects.filter(informe=instancia)
        #                     detalles_informe.update(status=False)
        #
        #                     instancia.niveltitulacion = nivel
        #                     instancia.institucioneducativa = institucion
        #                     instancia.descripcionmotivoconsulta = form.cleaned_data['descripcionmotivoconsulta']
        #                     instancia.motivoconsulta = motivocita
        #                     instancia.grado = int(request.POST['grado'])
        #                     instancia.save(request)
        #
        #                 # Procesar los datos de repuestaproceso y guardar en DetalleHistorialCitas
        #                 for respuesta in repuestaproceso:
        #                     idproceso = int(respuesta['id_proceso'])
        #                     observacion = ''
        #                     if 'observacion' in respuesta:
        #                         observacion = respuesta['observacion']
        #                     marcado = respuesta['marcado']
        #                 # Crear una instancia de DetalleHistorialCitas y asociarla a la instancia de HistorialClinicoCitas creada anteriormente
        #                     detalle_historial = DetalleHistorialPsicologico(informe=instancia, proceso_id=idproceso,
        #                                                               observacion=observacion, marcada=marcado)
        #                     detalle_historial.save()
        #
        #                 for respuesta in repuestadetalle:
        #                     estruc = EstructuraInforme.objects.get(id = int(respuesta['id']))
        #                     detalle = respuesta['respuesta']
        #                     detalle_estructura = DetalleInformePsicologico(descripcion=detalle, estructura=estruc, informe = instancia )
        #                     detalle_estructura.save()
        #
        #                 if 'archivo' in request.FILES:
        #                     newfile = request.FILES['archivo']
        #                     extension = newfile._name.split('.')
        #                     tam = len(extension)
        #                     exte = extension[tam - 1]
        #                     if newfile.size > 4194304:
        #                         transaction.set_rollback(True)
        #                         return JsonResponse(
        #                             {"result": True, "mensaje": u"Error, el tamaño del archivo es mayor a 2 Mb."})
        #                     if not exte.lower() in ['pdf']:
        #                         transaction.set_rollback(True)
        #                         return JsonResponse({"result": True, "mensaje": u"Error, solo archivos .pdf"})
        #                     newfile._name = generar_nombre(
        #                         "Respaldo_observacion_{}_{}".format(instancia.pk, random.randint(1, 100000).__str__()),
        #                         newfile._name)
        #                     instancia.archivo = newfile
        #                     instancia.save(request)
        #
        #                     log(u'Adiciono Historial de Citas: %s' % instancia, request, "add")
        #                 return JsonResponse(
        #                     {'result': False, 'mensaje': 'Guardado con éxito'})  # Agrega un mensaje de éxito
        #
        #             else:
        #                 transaction.set_rollback(True)
        #                 return JsonResponse(
        #                     {'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
        #                      "mensaje": "Error en el formulario"})
        #
        #         except Exception as ex:
        #             print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
        #             transaction.set_rollback(True)
        #             return JsonResponse({'result': True, "mensaje": 'Error: {}'.format(ex)}, safe=False)

        if action == 'addrefuerzopedagogico2':
            with transaction.atomic():
                try:
                    id = int(encrypt(request.POST['idpersonacita']))
                    # ids = int(encrypt(request.POST['ids']))
                    form = RefuerzoAcademicoForm(request.POST)
                    if form.is_valid():
                        if id != 0:
                            instance = PersonaCitaAgendada.objects.get(pk=id)
                            refuerzo = RefuerzoAcademico(personacita=instance,
                                                         destreza=form.cleaned_data['destreza'],
                                                         actividad=form.cleaned_data['actividad'],
                                                         fecha=hoy,
                                                         observacion=form.cleaned_data['observacion'],
                                                         asignatura = form.cleaned_data['asignatura'],
                                                         grado_egb = form.cleaned_data['grado_egb']
                                                         )
                            refuerzo.save(request)
                            if 'archivo' in request.FILES:
                                newfile = request.FILES['archivo']
                                extension = newfile._name.split('.')
                                tam = len(extension)
                                exte = extension[tam - 1]
                                if newfile.size > 4194304:
                                    transaction.set_rollback(True)
                                    return JsonResponse(
                                        {"result": True, "mensaje": u"Error, el tamaño del archivo es mayor a 2 Mb."})
                                if not exte.lower() in ['pdf']:
                                    transaction.set_rollback(True)
                                    return JsonResponse({"result": True, "mensaje": u"Error, solo archivos .pdf"})
                                newfile._name = generar_nombre(
                                    "Test_Aplicado_{}_{}".format(refuerzo.pk,
                                                                            random.randint(1, 100000).__str__()),
                                    newfile._name)
                                refuerzo.archivo = newfile
                                refuerzo.save(request)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                             "mensaje": "Error en el formulario"})
                    log(u'Adiciono refuerzo academico: %s' % instance, request, "add")
                    return JsonResponse({"result": False}, safe=False)
                except Exception as ex:
                    print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
                    transaction.set_rollback(True)
                    return JsonResponse({"result": True, "mensaje": "Intentelo más tarde: {}".format(str(ex))}, safe=False)

        if action == 'addrefuerzopedagogico':
            with transaction.atomic():
                try:
                    id = int(encrypt(request.POST['idpersonacita']))
                    # ids = int(encrypt(request.POST['ids']))
                    form = RefuerzoAcademicoForm(request.POST)
                    if form.is_valid():
                        if id != 0:
                            instance = PersonaCitaAgendada.objects.get(pk=id)
                            # refuerzocab = CabRefuerzoAcademico(personacita=instance,
                            #                                    asignatura=form.cleaned_data['asignatura'],
                            #                                    grado_egb=form.cleaned_data['grado_egb'])
                            refuerzocab = CabRefuerzoAcademico(
                                                               asignatura=form.cleaned_data['asignatura'],
                                                               grado_egb=form.cleaned_data['grado_egb'])
                            refuerzocab.save(request)

                            refuerzo = DetRefuerzoAcademico(personacita=instance, cabrefuerzo=refuerzocab,
                                                            destreza=form.cleaned_data['destreza'],
                                                            actividad=form.cleaned_data['actividad'],
                                                            fecha=hoy,
                                                            observacion=form.cleaned_data['observacion'])

                            refuerzo.save(request)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                             "mensaje": "Error en el formulario"})
                    log(u'Adiciono refuerzo academico: %s' % instance, request, "add")
                    return JsonResponse({"result": False}, safe=False)
                except Exception as ex:
                    print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
                    transaction.set_rollback(True)
                    return JsonResponse({"result": True, "mensaje": "Intentelo más tarde: {}".format(str(ex))}, safe=False)

        if action == 'editrefuerzopedagogico2':
            with transaction.atomic():
                try:
                    form = GestionServicioCitaForm(request.POST, request.FILES)
                    id = int(encrypt(request.POST['id']))
                    if form.is_valid():

                        refuerzo = RefuerzoAcademico.objects.get(id = int(encrypt(request.POST['id'])))
                        refuerzo.asignatura = request.POST['asignatura']
                        refuerzo.grado_egb = request.POST['grado_egb']
                        refuerzo.destreza = request.POST['destreza']
                        refuerzo.actividad = request.POST['actividad']
                        refuerzo.observacion = request.POST['observacion']
                        refuerzo.fecha = now()

                        refuerzo.save(request)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                             "mensaje": "Error en el formulario"})
                    log(u'Edito refuerzo academico %s' % refuerzo, request, "edit")
                    return JsonResponse({"result": False}, safe=False)
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": True, "mensaje": "Intentelo más tarde: {}".format(str(ex))}, safe=False)

        if action == 'editrefuerzopedagogico':
            with transaction.atomic():
                try:
                    form = GestionServicioCitaForm(request.POST, request.FILES)
                    id = int(encrypt(request.POST['id']))
                    if form.is_valid():

                        cab = CabRefuerzoAcademico.objects.get(id = int(encrypt(request.POST['idcab'])))
                        cab.asignatura = request.POST['asignatura']
                        cab.grado_egb = request.POST['grado_egb']
                        cab.save(request)

                        det = DetRefuerzoAcademico.objects.get(cabrefuerzo = cab)
                        det.destreza = request.POST['destreza']
                        det.actividad = request.POST['actividad']
                        det.observacion = request.POST['observacion']
                        det.fecha = now()
                        det.save(request)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                             "mensaje": "Error en el formulario"})
                    log(u'Edito refuerzo academico %s' % det, request, "edit")
                    return JsonResponse({"result": False}, safe=False)
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": True, "mensaje": "Intentelo más tarde: {}".format(str(ex))}, safe=False)

        if action == 'delrefuerzoacademico':
            with transaction.atomic():
                try:
                    refuerzo = RefuerzoAcademico.objects.get(id=int(encrypt(request.POST['id'])))
                    refuerzo.status=False
                    refuerzo.save(request)
                    log(u'Elimino refuerzo academico: %s' % refuerzo, request, "delrefuerzo")
                    res_json = {"error": False}
                except Exception as ex:
                    res_json = {'error': True, "message": "Error: {}".format(ex)}
                    return JsonResponse(res_json, safe=False)


        # if action == 'editinfomepsicologia':
        #     with transaction.atomic():
        #         try:
        #             data['readonly'] = False
        #             repuestaproceso = json.loads(request.POST['lista_items1'])
        #             repuestadetalle = repuestaproceso[-1]  # obtenemos solo el ultimo elemento
        #             repuestaproceso = repuestaproceso[:-1]  # obviamos el ultimo elemento
        #
        #             cita = PersonaCitaAgendada.objects.get(id=int(request.POST['persona']))
        #             form = InformePsicologicoForm(request.POST)
        #             if form.is_valid():
        #                 historial=InformePsicologico.objects.get(id=int(request.POST['ide']))
        #                 motivocita = MotivoCita.objects.get(id=int(request.POST['motivoconsulta']))
        #                 historial.motivoconsulta=motivocita
        #                 historial.descripcionmotivoconsulta = form.cleaned_data['descripcionmotivoconsulta']
        #                 historial.niveltitulacion = form.cleaned_data['niveltitulacion']
        #                 #historial.archivo = form.cleaned_data['archivo']
        #                 historial.save(request)
        #
        #                 detalles = DetalleHistorialPsicologico.objects.filter(historial__id=historial.id)
        #
        #                 for respuesta in repuestaproceso:
        #                     idp = respuesta['id_proceso']
        #                     detalle = detalles.filter(proceso__id=idp).first()
        #                     detalle.observacion = respuesta['observacion']
        #                     detalle.marcada = respuesta['marcado']
        #                     detalle.save()
        #
        #                 for respuesta in repuestadetalle:
        #                     estruc = EstructuraInforme.objects.get(id=int(respuesta['id']))
        #                     detalle = respuesta['respuesta']
        #                     detalle_estructura = DetalleInformePsicologico(descripcion=detalle, estructura=estruc)
        #                     detalle_estructura.save()
        #
        #                 if 'archivo' in request.FILES:
        #                     newfile = request.FILES['archivo']
        #                     extension = newfile._name.split('.')
        #                     tam = len(extension)
        #                     exte = extension[tam - 1]
        #                     if newfile.size > 4194304:
        #                         transaction.set_rollback(True)
        #                         return JsonResponse(
        #                             {"result": True, "mensaje": u"Error, el tamaño del archivo es mayor a 2 Mb."})
        #                     if not exte.lower() in ['pdf']:
        #                         transaction.set_rollback(True)
        #                         return JsonResponse({"result": True, "mensaje": u"Error, solo archivos .pdf"})
        #                     newfile._name = generar_nombre(
        #                         "Respaldo_observacion_{}_{}".format(historial.pk, random.randint(1, 100000).__str__()),
        #                         newfile._name)
        #                     historial.archivo = newfile
        #                     historial.save(request)
        #
        #                 log(u'Actualizo Historial de Citas: %s' % historial, request, "edit")
        #                 #fechas_historiales = HistorialClinicoCitas.objects.filter(persona=cita.persona).values_list('fechacita', flat=True)
        #                 return JsonResponse({"result": False, "mensaje": "Guardado correctamente"
        #                                      }, safe=False)
        #
        #         except Exception as ex:
        #             transaction.set_rollback(True)
        #             return JsonResponse({'result': True, "mensaje": 'Error: {}'.format(ex)}, safe=False)
        #

        # Datos familiares
        elif action == 'addfamiliar':
            try:
                id = encrypt_id(request.POST['idp'])
                persona = Persona.objects.get(id=id)
                f = FamiliarForm(request.POST)
                if f.is_valid():
                    edit_d = eval(request.POST.get('edit_d', ''))
                    cedula = f.cleaned_data['identificacion'].strip()
                    if persona.personadatosfamiliares_set.filter(identificacion=f.cleaned_data['identificacion'], status=True).exists():
                        raise NameError('El familiar se encuentra registrado.')
                    nombres = f"{f.cleaned_data['apellido1']} {f.cleaned_data['apellido2']} {f.cleaned_data['nombre']}"
                    familiar = PersonaDatosFamiliares(persona=persona,
                                                      identificacion=cedula,
                                                      nombre=nombres,
                                                      fallecido=f.cleaned_data['fallecido'],
                                                      nacimiento=f.cleaned_data['nacimiento'],
                                                      parentesco=f.cleaned_data['parentesco'],
                                                      tienediscapacidad=f.cleaned_data['tienediscapacidad'],
                                                      telefono=f.cleaned_data['telefono'],
                                                      telefono_conv=f.cleaned_data['telefono_conv'],
                                                      niveltitulacion=f.cleaned_data['niveltitulacion'],
                                                      ingresomensual=f.cleaned_data['ingresomensual'],
                                                      formatrabajo=f.cleaned_data['formatrabajo'],
                                                      trabajo=f.cleaned_data['trabajo'],
                                                      convive=f.cleaned_data['convive'],
                                                      sustentohogar=f.cleaned_data['sustentohogar'],
                                                      # rangoedad=f.cleaned_data['rangoedad'],
                                                      essustituto=f.cleaned_data['essustituto'],
                                                      autorizadoministerio=f.cleaned_data['autorizadoministerio'],
                                                      tipodiscapacidad=f.cleaned_data['tipodiscapacidad'],
                                                      porcientodiscapacidad=f.cleaned_data['porcientodiscapacidad'],
                                                      carnetdiscapacidad=f.cleaned_data['carnetdiscapacidad'],
                                                      institucionvalida=f.cleaned_data['institucionvalida'],
                                                      tipoinstitucionlaboral=f.cleaned_data['tipoinstitucionlaboral'],
                                                      negocio=f.cleaned_data['negocio'],
                                                      esservidorpublico=f.cleaned_data['esservidorpublico'],
                                                      bajocustodia=f.cleaned_data['bajocustodia'],
                                                      centrocuidado=f.cleaned_data['centrocuidado'] if f.cleaned_data['centrocuidado'] else 0,
                                                      centrocuidadodesc=f.cleaned_data['centrocuidadodesc'],
                                                      tienenegocio=f.cleaned_data['tienenegocio'])
                    familiar.save(request)
                    if 'cedulaidentidad' in request.FILES:
                        newfile = request.FILES['cedulaidentidad']
                        newfile._name = generar_nombre("cedulaidentidad_", newfile._name)
                        familiar.cedulaidentidad = newfile
                        familiar.save(request)
                    if 'ceduladiscapacidad' in request.FILES:
                        newfile = request.FILES['ceduladiscapacidad']
                        newfile._name = generar_nombre("ceduladiscapacidad_", newfile._name)
                        familiar.ceduladiscapacidad = newfile
                        familiar.save(request)
                    if 'archivoautorizado' in request.FILES:
                        newfile = request.FILES['archivoautorizado']
                        newfile._name = generar_nombre("archivoautorizado_", newfile._name)
                        familiar.archivoautorizado = newfile
                        familiar.save(request)
                    if 'cartaconsentimiento' in request.FILES:
                        newfile = request.FILES['cartaconsentimiento']
                        newfile._name = generar_nombre(f"cartaconsentimiento_{persona.usuario.username}", newfile._name)
                        familiar.cartaconsentimiento = newfile
                        familiar.save(request)
                    if 'archivocustodia' in request.FILES:
                        newfile = request.FILES['archivocustodia']
                        newfile._name = generar_nombre(f"archivocustodia_{persona.usuario.username}", newfile._name)
                        familiar.archivocustodia = newfile
                        familiar.save(request)

                    pers = Persona.objects.filter(Q(pasaporte=cedula) | Q(cedula=cedula) | Q(pasaporte=('VS' + cedula)) | Q(cedula=cedula[2:]),status=True).first()
                    if not pers:
                        pers = Persona(cedula=f.cleaned_data['identificacion'],
                                       nombres=f.cleaned_data['nombre'],
                                       apellido1=f.cleaned_data['apellido1'],
                                       apellido2=f.cleaned_data['apellido2'],
                                       nacimiento=f.cleaned_data['nacimiento'],
                                       telefono=f.cleaned_data['telefono'],
                                       sexo=f.cleaned_data['sexo'],
                                       telefono_conv=f.cleaned_data['telefono_conv'],
                                       )
                        pers.save(request)
                        log(u'Adiciono persona: %s' % persona, request, "add")
                    elif len(pers.mis_perfilesusuarios()) == 1 and pers.tiene_usuario_externo():
                        pers.cedula = f.cleaned_data['identificacion']
                        pers.nombres = f.cleaned_data['nombre']
                        pers.apellido1 = f.cleaned_data['apellido1']
                        pers.apellido2 = f.cleaned_data['apellido2']
                        pers.nacimiento = f.cleaned_data['nacimiento']
                        pers.telefono = f.cleaned_data['telefono']
                        pers.sexo = f.cleaned_data['sexo']
                        pers.telefono_conv = f.cleaned_data['telefono_conv']
                        pers.save(request)
                        log(u'Edito familiar de usuario: %s' % pers, request, "edit")
                    if not pers.tiene_perfil():
                        if not Externo.objects.filter(persona=pers, status=True):
                            externo = Externo(persona=pers)
                            externo.save(request)
                            log(u'Adiciono externo: %s' % pers, request, "add")
                        perfil = PerfilUsuario(persona=pers, externo=externo)
                        perfil.save(request)
                        log(u'Adiciono perfil de usuario: %s' % perfil, request, "add")

                    perfil_i=pers.mi_perfil()
                    if not edit_d and perfil_i.tienediscapacidad:
                        familiar.tienediscapacidad=perfil_i.tienediscapacidad
                        familiar.tipodiscapacidad=perfil_i.tipodiscapacidad
                        familiar.porcientodiscapacidad=perfil_i.porcientodiscapacidad
                        familiar.carnetdiscapacidad=perfil_i.carnetdiscapacidad
                        familiar.institucionvalida=perfil_i.institucionvalida
                        familiar.ceduladiscapacidad=perfil_i.archivo.name if perfil_i.archivo else ''
                        familiar.archivoautorizado=perfil_i.archivovaloracion.name if perfil_i.archivovaloracion else ''
                    elif f.cleaned_data['tienediscapacidad']:
                        perfil_i.tienediscapacidad = f.cleaned_data['tienediscapacidad']
                        perfil_i.tipodiscapacidad = f.cleaned_data['tipodiscapacidad']
                        perfil_i.porcientodiscapacidad = f.cleaned_data['porcientodiscapacidad'] if f.cleaned_data['porcientodiscapacidad'] else 0
                        perfil_i.carnetdiscapacidad = f.cleaned_data['carnetdiscapacidad']
                        perfil_i.institucionvalida = f.cleaned_data['institucionvalida']
                        if 'ceduladiscapacidad' in request.FILES:
                            newfile = request.FILES['ceduladiscapacidad']
                            newfile._name = generar_nombre("archivosdiscapacidad_", newfile._name)
                            perfil_i.archivo = newfile
                            perfil_i.estadoarchivodiscapacidad = 1
                        if 'archivoautorizado' in request.FILES:
                            newfile = request.FILES['archivoautorizado']
                            newfile._name = generar_nombre("archivovaloracionmedica_", newfile._name)
                            perfil_i.archivovaloracion = newfile
                        perfil_i.save(request)
                    if f.cleaned_data['parentesco'].id in [14,11] and not persona.apellido1 in [pers.apellido1,pers.apellido2]:
                        familiar.aprobado=False
                    if f.cleaned_data['parentesco'].id == 13 and not pers.personadatosfamiliares_set.filter(personafamiliar=persona, status=True).exists():
                        fam_ = PersonaDatosFamiliares(persona=pers,
                                                      personafamiliar=persona,
                                                      identificacion=persona.cedula,
                                                      nombre=persona.nombre_completo_inverso(),
                                                      nacimiento=persona.nacimiento,
                                                      parentesco=f.cleaned_data['parentesco'],
                                                      telefono=persona.telefono,
                                                      telefono_conv=persona.telefono_conv,
                                                      convive=f.cleaned_data['convive'])
                        fam_.save(request)
                        perfil_fam = persona.mi_perfil()
                        if perfil_fam.tienediscapacidad:
                            fam_.tienediscapacidad = perfil_fam.tienediscapacidad
                            fam_.tipodiscapacidad = perfil_fam.tipodiscapacidad
                            fam_.porcientodiscapacidad = perfil_fam.porcientodiscapacidad
                            fam_.carnetdiscapacidad = perfil_fam.carnetdiscapacidad
                            fam_.institucionvalida = perfil_fam.institucionvalida
                            fam_.ceduladiscapacidad = perfil_fam.archivo.name if perfil_i.archivo else ''
                            fam_.archivoautorizado = perfil_fam.archivovaloracion.name if perfil_i.archivovaloracion else ''
                        fam_.save(request)
                    familiar.personafamiliar=pers
                    familiar.save(request)
                    if familiar.parentesco.id in [11, 14] or familiar.bajocustodia:
                        per_extension = persona.personaextension_set.filter(status=True).first()
                        hijos = per_extension.hijos if per_extension.hijos else 0
                        per_extension.hijos = hijos+1
                        per_extension.save(request)
                    log(u'Adiciono familiar: %s' % familiar, request, "add")
                    return JsonResponse({'result': False, 'mensaje':'Guardado con exito'})
                else:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "mensaje": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': f'{ex}'})

        elif action == 'editfamiliar':
            try:
                persona = request.session['persona']
                f = FamiliarForm(request.POST)
                f.edit()
                if f.is_valid():
                    familiar = PersonaDatosFamiliares.objects.get(pk=encrypt_id(request.POST['id']))
                    edit_d=eval(request.POST.get('edit_d',''))
                    cedula = f.cleaned_data['identificacion'].strip()
                    if persona.personadatosfamiliares_set.filter(identificacion=cedula, status=True).exclude(id=familiar.id).exists():
                        raise NameError(u'El familiar se encuentra registrado.')
                    nombres = f"{f.cleaned_data['apellido1']} {f.cleaned_data['apellido2']} {f.cleaned_data['nombre']}"
                    pers = Persona.objects.filter(Q(pasaporte=cedula) | Q(cedula=cedula) | Q(pasaporte=('VS' + cedula)) | Q(cedula=cedula[2:]),status=True).first()
                    familiar.fallecido = f.cleaned_data['fallecido']
                    familiar.parentesco = f.cleaned_data['parentesco']
                    familiar.tienediscapacidad = f.cleaned_data['tienediscapacidad']
                    familiar.trabajo = f.cleaned_data['trabajo']
                    familiar.niveltitulacion = f.cleaned_data['niveltitulacion']
                    familiar.ingresomensual = f.cleaned_data['ingresomensual']
                    familiar.formatrabajo = f.cleaned_data['formatrabajo']
                    familiar.convive = f.cleaned_data['convive']
                    familiar.sustentohogar = f.cleaned_data['sustentohogar']
                    # familiar.rangoedad = f.cleaned_data['rangoedad']
                    familiar.tienenegocio = f.cleaned_data['tienenegocio']
                    familiar.esservidorpublico = f.cleaned_data['esservidorpublico']
                    familiar.bajocustodia = f.cleaned_data['bajocustodia']
                    familiar.centrocuidado = f.cleaned_data['centrocuidado'] if f.cleaned_data['centrocuidado'] else 0
                    familiar.centrocuidadodesc = f.cleaned_data['centrocuidadodesc']
                    familiar.negocio = ''
                    familiar.tipoinstitucionlaboral = f.cleaned_data['tipoinstitucionlaboral']
                    if familiar.tienenegocio:
                        familiar.negocio = f.cleaned_data['negocio']
                    familiar.save(request)
                    if 'cedulaidentidad' in request.FILES:
                        newfile = request.FILES['cedulaidentidad']
                        newfile._name = generar_nombre("cedulaidentidad_", newfile._name)
                        familiar.cedulaidentidad = newfile
                        familiar.save(request)
                    if 'ceduladiscapacidad' in request.FILES:
                        newfile = request.FILES['ceduladiscapacidad']
                        newfile._name = generar_nombre("ceduladiscapacidad_", newfile._name)
                        familiar.ceduladiscapacidad = newfile
                        familiar.save(request)
                    if 'cartaconsentimiento' in request.FILES:
                        newfile = request.FILES['cartaconsentimiento']
                        newfile._name = generar_nombre(f"cartaconsentimiento_{persona.usuario.username}", newfile._name)
                        familiar.cartaconsentimiento = newfile
                        familiar.save(request)
                    if 'archivocustodia' in request.FILES:
                        newfile = request.FILES['archivocustodia']
                        newfile._name = generar_nombre(f"archivocustodia_{persona.usuario.username}", newfile._name)
                        familiar.archivocustodia = newfile
                        familiar.save(request)

                    if f.cleaned_data['tienediscapacidad']:
                        if 'archivoautorizado' in request.FILES:
                            newfile = request.FILES['archivoautorizado']
                            newfile._name = generar_nombre("archivoautorizado_", newfile._name)
                            familiar.archivoautorizado = newfile
                            familiar.save(request)
                    else:
                        familiar.archivoautorizado = None
                        familiar.save(request)
                    if not pers:
                        pers = Persona(cedula=f.cleaned_data['identificacion'],
                                       nombres=f.cleaned_data['nombre'],
                                       apellido1=f.cleaned_data['apellido1'],
                                       apellido2=f.cleaned_data['apellido2'],
                                       nacimiento=f.cleaned_data['nacimiento'],
                                       telefono=f.cleaned_data['telefono'],
                                       sexo=f.cleaned_data['sexo'],
                                       telefono_conv=f.cleaned_data['telefono_conv'],
                                       )
                        pers.save(request)
                        log(u'Adiciono persona: %s' % persona, request, "add")
                    elif len(pers.mis_perfilesusuarios()) == 1 and pers.tiene_usuario_externo():
                        pers.cedula = cedula
                        pers.nombres = f.cleaned_data['nombre']
                        pers.apellido1 = f.cleaned_data['apellido1']
                        pers.apellido2 = f.cleaned_data['apellido2']
                        pers.nacimiento = f.cleaned_data['nacimiento']
                        pers.telefono = f.cleaned_data['telefono']
                        pers.sexo = f.cleaned_data['sexo']
                        pers.telefono_conv = f.cleaned_data['telefono_conv']
                        pers.save(request)
                        log(u'Edito familiar de usuario: %s' % pers, request, "edit")
                    if not pers.tiene_perfil():
                        if not Externo.objects.filter(persona=pers, status=True):
                            externo = Externo(persona=pers)
                            externo.save(request)
                            log(u'Adiciono externo: %s' % pers, request, "add")
                        perfil = PerfilUsuario(persona=pers, externo=externo)
                        perfil.save(request)
                        log(u'Adiciono perfil de usuario: %s' % perfil, request, "add")
                    if len(pers.mis_perfilesusuarios()) == 1 and pers.tiene_usuario_externo():
                        familiar.identificacion = cedula
                        familiar.nombre = nombres
                        familiar.nacimiento = f.cleaned_data['nacimiento']
                        familiar.telefono = f.cleaned_data['telefono']
                        familiar.telefono_conv = f.cleaned_data['telefono_conv']
                    else:
                        familiar.identificacion = pers.cedula
                        familiar.nombre = pers.nombre_completo_inverso()
                        familiar.nacimiento = pers.nacimiento
                        familiar.telefono = pers.telefono
                        familiar.telefono_conv = pers.telefono_conv
                    perfil_i = pers.mi_perfil()
                    if edit_d:
                        if f.cleaned_data['tienediscapacidad']:
                            familiar.essustituto = f.cleaned_data['essustituto']
                            familiar.autorizadoministerio = f.cleaned_data['autorizadoministerio']
                            familiar.tipodiscapacidad = f.cleaned_data['tipodiscapacidad']
                            familiar.porcientodiscapacidad = f.cleaned_data['porcientodiscapacidad']
                            familiar.carnetdiscapacidad = f.cleaned_data['carnetdiscapacidad']
                            familiar.institucionvalida = f.cleaned_data['institucionvalida']
                            perfil_i.tienediscapacidad = f.cleaned_data['tienediscapacidad']
                            perfil_i.tipodiscapacidad = f.cleaned_data['tipodiscapacidad']
                            perfil_i.porcientodiscapacidad = f.cleaned_data['porcientodiscapacidad'] if f.cleaned_data['porcientodiscapacidad'] else 0
                            perfil_i.carnetdiscapacidad = f.cleaned_data['carnetdiscapacidad']
                            perfil_i.institucionvalida = f.cleaned_data['institucionvalida']
                            if 'ceduladiscapacidad' in request.FILES:
                                newfile = request.FILES['ceduladiscapacidad']
                                newfile._name = generar_nombre("archivosdiscapacidad_", newfile._name)
                                perfil_i.archivo = newfile
                                perfil_i.estadoarchivodiscapacidad = 1
                            if 'archivoautorizado' in request.FILES:
                                newfile = request.FILES['archivoautorizado']
                                newfile._name = generar_nombre("archivovaloracionmedica_", newfile._name)
                                perfil_i.archivovaloracion = newfile
                            perfil_i.save(request)
                        else:
                            familiar.essustituto = False
                            familiar.autorizadoministerio = False
                            familiar.tipodiscapacidad = None
                            familiar.porcientodiscapacidad = None
                            familiar.carnetdiscapacidad = ''
                            familiar.institucionvalida = None
                            perfil_i.tienediscapacidad = False
                            perfil_i.tipodiscapacidad = None
                            perfil_i.porcientodiscapacidad = None
                            perfil_i.carnetdiscapacidad = ''
                            perfil_i.institucionvalida = None
                            perfil_i.save(request)
                    elif perfil_i.tienediscapacidad:
                        familiar.essustituto = f.cleaned_data['essustituto']
                        familiar.autorizadoministerio = f.cleaned_data['autorizadoministerio']
                        familiar.tienediscapacidad = perfil_i.tienediscapacidad
                        familiar.tipodiscapacidad = perfil_i.tipodiscapacidad
                        familiar.porcientodiscapacidad = perfil_i.porcientodiscapacidad
                        familiar.carnetdiscapacidad = perfil_i.carnetdiscapacidad
                        familiar.institucionvalida = perfil_i.institucionvalida
                        familiar.ceduladiscapacidad = perfil_i.archivo.name if perfil_i.archivo else ''
                        familiar.archivoautorizado = perfil_i.archivovaloracion.name if perfil_i.archivovaloracion else ''
                        perfil_i.save(request)
                    if f.cleaned_data['parentesco'].id in [14, 11] and not persona.apellido1 in [pers.apellido1, pers.apellido2]:
                        familiar.aprobado = False
                    if f.cleaned_data['parentesco'].id == 13 and not pers.personadatosfamiliares_set.filter(personafamiliar=persona, status=True).exists():
                        fam_ = PersonaDatosFamiliares(persona=pers,
                                                      personafamiliar=persona,
                                                      identificacion=persona.cedula,
                                                      nombre=persona.nombre_completo_inverso(),
                                                      nacimiento=persona.nacimiento,
                                                      parentesco=f.cleaned_data['parentesco'],
                                                      telefono=persona.telefono,
                                                      telefono_conv=persona.telefono_conv,
                                                      convive=f.cleaned_data['convive'])
                        fam_.save(request)
                        perfil_fam = persona.mi_perfil()
                        if perfil_fam.tienediscapacidad:
                            fam_.tienediscapacidad = perfil_fam.tienediscapacidad
                            fam_.tipodiscapacidad = perfil_fam.tipodiscapacidad
                            fam_.porcientodiscapacidad = perfil_fam.porcientodiscapacidad
                            fam_.carnetdiscapacidad = perfil_fam.carnetdiscapacidad
                            fam_.institucionvalida = perfil_fam.institucionvalida
                            fam_.ceduladiscapacidad = perfil_fam.archivo.name if perfil_i.archivo else ''
                            fam_.archivoautorizado = perfil_fam.archivovaloracion.name if perfil_i.archivovaloracion else ''
                        fam_.save(request)
                    familiar.personafamiliar = pers
                    familiar.save(request)
                    log(u'Modifico familiar: %s' % persona, request, "edit")
                    return JsonResponse({'result': False, 'mensaje': 'Guardado con exito'})
                else:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "mensaje": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': str(ex)})


        return JsonResponse({"result": "bad", "mensaje": u"Solicitud Incorrecta."})
    else:
        if 'action' in request.GET:
            data['action'] = action = request.GET['action']
            # Gestionar cita principal
            if action == 'finalizar':
                try:
                    if request.GET['id']:
                        data['cita'] = instance = PersonaCitaAgendada.objects.get(pk=int(request.GET['id']))
                    elif request.GET['ids']:
                        data['subcita'] = instance = SubCitaAgendada.objects.get(pk=int(request.GET['ids']))
                    form = FinalizaCitaForm(initial=model_to_dict(instance))
                    data['form'] = form
                    template = get_template("adm_agendamientocitas/modal/formfinalizar.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'validar':
                try:
                    form = ValidarCitaForm()
                    data['form'] = form
                    data['cita'] = PersonaCitaAgendada.objects.get(pk=request.GET['id'])
                    data['estados'] = ESTADOS_DOCUMENTOS_SOLICITUD[0:]
                    template = get_template("adm_agendamientocitas/modal/formvalidar.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass


            elif action == 'refuerzopedagogico2':
                try:
                    data['title'] = u'Gestionar Refuerzo Académico'
                    cita_id = int(encrypt(request.GET['id']))
                    data['cita'] = cita = PersonaCitaAgendada.objects.get(id=cita_id)
                    PersonaCitaAgendada.objects.get(id=int(encrypt(request.GET['id'])))
                    data['tipoformulario'] = tipoformulario = int(request.GET['tipoformulario'])
                    tipo_informe, id_serviciocita = tipo_informe_servicio(cita, persona)
                    if not tipo_informe:
                        raise NameError('Usted no tiene acceso a este apartado')
                    serviciopersona = cita.servicio_persona(persona)
                    asignados = asignados_id(cita)
                    data['asignados'] = asignados

                    data['refuerzo'] = refuerzo = RefuerzoAcademico.objects.filter(personacita=cita, status=True)
                    ultimo = RefuerzoAcademico.objects.filter(personacita=cita, status=True).order_by('-id').first()
                    if ultimo:
                        dato = RefuerzoAcademico.objects.filter(id=ultimo.id).first()
                        if dato:
                            anio = RefuerzoAcademico.objects.get(id=ultimo.id)
                            data['anio'] = anio
                        else :
                            data['anio'] = ''
                    return render(request, 'adm_agendamientocitas/viewrefuerzoacademico.html', data)
                except Exception as ex:
                    print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
                    messages.error(request, f'{ex}')

            elif action == 'gestionarcita':
                try:
                    data['title'] = u'Gestionar cita'
                    data['cita'] = cita = PersonaCitaAgendada.objects.get(id=int(encrypt(request.GET['id'])))
                    data['Serviciocitaid'] = cita.servicio.serviciocita.departamentoservicio.id
                    data['asignados'] =asignados_id(cita)
                    data['navactivo'] = 'observaciones'
                    data['responsableid'] = persona.id
                    if not persona.id in cita.responsables_subcitas_list() and not usuario.is_superuser:
                        raise NameError('Usted no tiene acceso a esta cita.')

                    data['estadocivil'] = ''
                    if cita.persona.datos_extension().estadocivil != None:
                        data['estadocivil'] = cita.persona.datos_extension().estadocivil.nombre

                    request.session['viewactivo'] = 1
                    return render(request, 'adm_agendamientocitas/viewgestorcita.html', data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            # elif action == 'mostrar_observaciones_derivado':
            #     try:
            #         lista = []
            #         if 'ids' in request.GET:
            #             id = int(request.GET['ids'])
            #             observaciones = GestionServicioCita.objects.filter(cita__departamentoservicio_id=id)
            #             for obs in observaciones:
            #                 lista.append(
            #                     {'id': encrypt(obs.id), 'observacion': obs.observacion, 'mostrar': obs.mostrar})
            #             return JsonResponse({'result': True, 'lista': lista})
            #         else:
            #             id = int(request.GET['id'])
            #             observaciones = GestionServicioCita.objects.filter(subcita__departamentoservicio_id=id)
            #             for obs in observaciones:
            #                 lista.append(
            #                     {'id': encrypt(obs.id), 'observacion': obs.observacion, 'mostrar': obs.mostrar})
            #             return JsonResponse({'result': True, 'data': lista})
            #     except Exception as e:
            #         return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})


            elif action == 'addrefuerzopedagogico2':
                try:
                    data['form'] = RefuerzoAcademicoForm()
                    if request.GET['id']:
                        # data['cita'] = cita = CabRefuerzoAcademico.objects.get(persona=cita)
                        data['cita'] = cita = PersonaCitaAgendada.objects.get(pk=int(encrypt(request.GET['id'])))

                        refuerzo = RefuerzoAcademico.objects.filter(personacita=cita).order_by('-id').first()
                        if refuerzo:
                            ultimo = RefuerzoAcademico.objects.get(id=refuerzo.id)
                            datos_combinados = {
                                'grado_egb': ultimo.grado_egb
                            }
                            data['form'] = RefuerzoAcademicoForm(initial=datos_combinados)

                    elif request.GET['ids']:
                        data['subcita'] = SubCitaAgendada.objects.get(pk=int(encrypt(request.GET['ids'])))
                    template = get_template("adm_agendamientocitas/modal/formrefuerzopedagogico.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
                    pass


            elif action == 'editrefuerzopedagogico2':
                try:
                    data['refuerzo'] = refuerzo = RefuerzoAcademico.objects.get(pk=int(encrypt(request.GET['id'])))

                    # if not cabrefuerzo:
                    #     cabrefuerzo = CabRefuerzoAcademico(personacita=refuerzo.cita)
                    #     cabrefuerzo.save()
                    # Pass the existing header information to the form
                    data['form'] = RefuerzoAcademicoForm(initial=model_to_dict(refuerzo))
                    template = get_template("adm_agendamientocitas/modal/formrefuerzopedagogico.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
                    pass

            elif action == 'addgestion':
                try:
                    data['form'] = GestionServicioCitaForm()
                    if request.GET['id']:
                        data['cita'] = PersonaCitaAgendada.objects.get(pk=int(encrypt(request.GET['id'])))
                    elif request.GET['ids']:
                        data['subcita'] = SubCitaAgendada.objects.get(pk=int(encrypt(request.GET['ids'])))
                    template = get_template("adm_agendamientocitas/modal/formgestioncita.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'editgestion':
                try:
                    data['gestion']= gestion= GestionServicioCita.objects.get(id=int(encrypt(request.GET['id'])))
                    data['form'] = GestionServicioCitaForm(initial=model_to_dict(gestion))
                    template = get_template("adm_agendamientocitas/modal/formgestioncita.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass
            # Gestionar cita principal
            elif action == 'gestionarsubcita':
                try:
                    data['title'] = u'Gestionar sub cita'
                    data['subcita_'] = subcita = SubCitaAgendada.objects.get(id=int(encrypt(request.GET['id'])))
                    data['cita']=subcita.citaprincipal
                    data['navactivo']='observaciones'
                    data['responsableid'] = persona.id
                    request.session['viewactivo'] = 2

                    return render(request, 'adm_agendamientocitas/viewgestorsubcita.html', data)
                except Exception as ex:
                    print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
                    pass
            # Gestionar cita principal

            # Agendar Subcita
            elif action == 'agendarcita':
                try:
                    data['title'] = 'Agendar Cita'
                    data['cita'] = cita = PersonaCitaAgendada.objects.get(pk=int(encrypt(request.GET['id'])))
                    esderivada = False
                    if cita.persona_responsable_id != persona.id:
                        esderivada = True
                    if esderivada:
                        ultimo = cita.ultima_subcita()
                        data['servicio'] = ultimo.servicio
                    else:
                        data['servicio'] = ServicioConfigurado.objects.get(pk=cita.servicio.id)
                    data['departamentos'] = DepartamentoServicio.objects.filter(status=True)
                    data['navactivo'] = 'observaciones'
                    request.session['viewactivo'] = 3
                    return render(request, 'adm_agendamientocitas/viewagendarcita.html', data)
                except Exception as ex:
                    pass

            elif action == 'cargarturno':
                try:
                    idresponsableservicio = request.GET['idpersona']
                    lista = json.loads(request.GET['listahorarios'])

                    fecha = request.GET['fecha']

                    horarios = HorarioServicioCita.objects.filter(id__in=lista,
                                                                  responsableservicio_id=idresponsableservicio,
                                                                  status=True)
                    horarios_disponibles = []

                    for horario in horarios:
                        turnos_disponibles = horario.citas_disponibles(fecha)
                        if turnos_disponibles > 0:
                            disponible = horario.horario_disponible(fecha)
                            if disponible:
                                horarios_disponibles.append({
                                    'id': horario.id,
                                    'nombre_horario': horario.turno.nombre_horario(),
                                    'id_tipo_atencion': horario.tipo_atencion,
                                    'tipo_atencion': horario.get_tipo_atencion_display(),
                                    'turnos': turnos_disponibles,
                                })

                    return JsonResponse({'result': 'ok', 'horarios': horarios_disponibles})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Solicitud Incorrecta."})

            elif action == 'cargarcalendario':
                try:
                    fecha = datetime.now().date()
                    panio = fecha.year
                    pmes = fecha.month
                    if 'mover' in request.GET:
                        mover = request.GET['mover']

                        if mover == 'anterior':
                            mes = int(request.GET['mes'])
                            anio = int(request.GET['anio'])
                            pmes = mes - 1
                            if pmes == 0:
                                pmes = 12
                                panio = anio - 1
                            else:
                                panio = anio

                        elif mover == 'proximo':
                            mes = int(request.GET['mes'])
                            anio = int(request.GET['anio'])
                            pmes = mes + 1
                            if pmes == 13:
                                pmes = 1
                                panio = anio + 1
                            else:
                                panio = anio

                    id = int(encrypt(request.GET['idservicio']))
                    data['idresponsable'] = idresponsable = request.GET.get('idresponsable', '')
                    idresponsable = int(idresponsable) if idresponsable else persona.id
                    s_anio = panio
                    s_mes = pmes
                    s_dia = 1
                    data['mes'] = MESES_CHOICES[s_mes - 1]
                    data['ws'] = [0, 7, 14, 21, 28, 35]
                    lista = {}
                    listahorarios = []
                    data['servicio'] = servicio = ServicioConfigurado.objects.get(pk=id)
                    # finicio = hoy - timedelta(days=servicio.numdias)
                    data['horarios'] = horarios = HorarioServicioCita.objects.filter(status=True,
                                                                                     responsableservicio__servicio_id=id,
                                                                                     responsableservicio__responsable_id=idresponsable,
                                                                                     mostrar=True,
                                                                                     fechafin__gte=hoy,
                                                                                     # fechainicio__gte=finicio
                                                                                     )
                    # diasreserva = servicio.numdias
                    # anio_actual = hoy.year
                    # if hoy.month != s_mes:
                    #     anioubicado=s_anio
                    #     cont=0
                    #     while anioubicado >= anio_actual:
                    #         iden = True
                    #         if anio_actual == anioubicado and cont == 0:
                    #             mes = hoy.month
                    #             mesubicado = s_mes
                    #         elif anio_actual == anioubicado:
                    #             mes = hoy.month
                    #             mesubicado = 12
                    #         elif anio_actual < anioubicado and cont == 0:
                    #             mes = 0
                    #             mesubicado = s_mes
                    #         elif anio_actual < anioubicado:
                    #             mes = 0
                    #             mesubicado = 12
                    #         while mesubicado > mes:
                    #             if anio_actual < anioubicado and iden:
                    #                 mes=1
                    #                 iden=False
                    #             numsinhorario = 0
                    #             msiguiente = calendar.monthrange(s_anio, mes)
                    #             rango = range(1, int(msiguiente[1] + 1), 1)
                    #             if mes == hoy.month:
                    #                 rango = range(int(hoy.day), int(msiguiente[1] + 1), 1)
                    #             for dia in rango:
                    #                 ban = False
                    #                 fecha = date(s_anio, mes, dia)
                    #                 numdia = fecha.weekday() + 1
                    #                 for horario in horarios:
                    #                     if horario.dia == numdia:
                    #                         ban = True
                    #                 if ban == False:
                    #                     numsinhorario += 1
                    #             diasreserva += numsinhorario
                    #             if mes == hoy.month:
                    #                 diasreserva -= msiguiente[1] - (hoy.day - 1)
                    #             else:
                    #                 diasreserva = diasreserva - msiguiente[1]
                    #
                    #             mes += 1
                    #         cont+=1
                    #         anioubicado-=1

                    for i in range(1, 43, 1):
                        dia = {i: 'no'}
                        lista.update(dia)
                    comienzo = False
                    fin = False
                    for i in lista.items():
                        try:
                            fecha = date(s_anio, s_mes, s_dia)
                            if fecha.isoweekday() == i[0] and fin is False and comienzo is False:
                                comienzo = True
                        except Exception as ex:
                            pass
                        if comienzo:
                            try:
                                fecha = date(s_anio, s_mes, s_dia)
                            except Exception as ex:
                                fin = True
                        if comienzo and fin is False:
                            dia = {i[0]: s_dia}
                            lista.update(dia)
                            listhorario = []
                            sinhorario = True
                            puedereservar = False
                            numerodia = fecha.weekday() + 1
                            if fecha >= hoy:
                                turnos = 0
                                listturnos = []
                                for horario in horarios:
                                    if horario.dia == numerodia and fecha <= horario.fechafin and fecha >= horario.fechainicio:
                                        listhorario.append(horario.id)
                                if listhorario:
                                    sinhorario = False
                                    filter = horarios.filter(status=True, pk__in=listhorario)
                                    for horario in filter:
                                        listturnos.append(horario.turno.comienza)
                                        turnos += horario.citas_disponibles(fecha)

                                monthRange = calendar.monthrange(s_anio, s_mes)
                                # if diasreserva > 0 and s_dia <= monthRange[1] and not sinhorario:
                                if s_dia <= monthRange[1] and not sinhorario:
                                    ordenadas = []
                                    ordenadas = sorted(listturnos, reverse=True)
                                    if ordenadas[0] > datetime.now().time() or fecha > datetime.now().date():
                                        puedereservar = True
                                        # diasreserva -= 1
                                diccionario = {'dia': s_dia, 'turnos': turnos, 'listahorario': listhorario,
                                               'sinhorario': sinhorario, 'fecha': fecha, 'puedereservar': puedereservar}
                                listahorarios.append(diccionario)
                            s_dia += 1

                    if horarios:
                        data['ultimafecha'] = ultimafecha = horarios.order_by('fechafin').last().fechafin
                        primerafecha = horarios.order_by('fechafin').first().fechafin
                        anio = primerafecha.year
                        rango = ultimafecha.year - primerafecha.year
                        for i in range(0, rango, 1):
                            if s_anio == anio + (i + 1):
                                anio = s_anio
                        data['year'] = anio
                    data['dias_mes'] = lista
                    data['dwnm'] = ['lunes', 'martes', 'miercoles', 'jueves', 'viernes', 'sabado', 'domingo']
                    data['dwn'] = [1, 2, 3, 4, 5, 6, 7]
                    data['daymonth'] = 1
                    data['s_anio'] = s_anio
                    data['s_mes'] = s_mes
                    data['lista'] = lista
                    data['hoy'] = hoy
                    data['hoy_dia'] = hoy.day
                    data['hoy_mes'] = hoy.month
                    data['listahorarios'] = listahorarios
                    ultimo_dia = calendar.monthrange(hoy.year, hoy.month)[1]
                    data['fechaactual']=date(hoy.year, hoy.month, ultimo_dia)
                    data['fechacalendario'] = date(s_anio, s_mes, s_dia - 1)
                    template = get_template("adm_agendamientocitas/viewcalendario.html")
                    return JsonResponse({"result": 'ok', 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": 'bad', 'mensaje': 'Error'.format(ex)})

            elif action == 'addcita':
                try:
                    servicio = ServicioConfigurado.objects.get(pk=int(encrypt(request.GET['idservicio'])))
                    citaprincipal = PersonaCitaAgendada.objects.get(pk= int(encrypt(request.GET['idcita'])))
                    horariosid = json.loads(request.GET["listaid[]"])
                    # if not type(eval(horariosid[0])) is int:
                    #     horariosid = list(eval(horariosid[0]))
                    data['fecha'] = request.GET['fecha']
                    data['servicio'] = servicio
                    data['cita'] = citaprincipal
                    data['horarios'] = horario = HorarioServicioCita.objects.filter(id__in=horariosid, status=True).order_by('id')
                    data['horariodia'] = horario.first()
                    template = get_template("adm_agendamientocitas/modal/formagendar.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'mensaje': 'Error'.format(ex)})
            # Agendar Subcita

            # Agendar cita por administrativos
            elif action == 'admagendar':

                try:
                    data['title']=u'Agendamiento de citas'
                    data['departamentos']=DepartamentoServicio.objects.filter(status=True)
                    return render(request, 'adm_agendamientocitas/viewadmagendarcitas.html', data)
                except Exception as ex:
                    pass

            elif action == 'buscarpersonas':
                try:
                    q = request.GET['q'].upper().strip()
                    s = q.split(" ")
                    qspersona = Persona.objects.filter(status=True).order_by('apellido1')
                    if len(s) == 1:
                        qspersona = qspersona.filter((Q(nombres__icontains=q) | Q(apellido1__icontains=q) | Q(
                            cedula__icontains=q) | Q(apellido2__icontains=q) | Q(cedula__contains=q)))[:15]
                    elif len(s) == 2:
                        qspersona = qspersona.filter((Q(apellido1__contains=s[0]) & Q(apellido2__contains=s[1])) |
                                                     (Q(nombres__icontains=s[0]) & Q(nombres__icontains=s[1])) |
                                                     (Q(nombres__icontains=s[0]) & Q(apellido1__contains=s[1]))).distinct()[:15]
                    else:
                        qspersona = qspersona.filter(
                            (Q(nombres__contains=s[0]) & Q(apellido1__contains=s[1]) & Q(apellido2__contains=s[2])) |
                            (Q(nombres__contains=s[0]) & Q(nombres__contains=s[1]) & Q(
                                apellido1__contains=s[2]))).distinct()[:15]

                    resp = [{'id': qs.pk, 'text': f"{qs.nombre_completo_inverso()}",
                             'documento': qs.documento(),
                             'foto': qs.get_foto()} for qs in qspersona]
                    return HttpResponse(json.dumps({'status': True, 'results': resp}))
                except Exception as ex:
                    pass

            elif action == 'listperfiles':
                try:
                    lista = []
                    idpersona = int(request.GET['idpersona'])
                    perfiles = PerfilUsuario.objects.filter(status=True, persona_id=idpersona, visible=True)
                    for p in perfiles:
                        text = str(p)
                        if p.inscripcion:
                            text = "{} ({})".format(str(p), str(p.inscripcion.carrera))
                        lista.append([p.id, text])
                    return JsonResponse({'result': 'ok', 'lista': lista})
                except Exception as e:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            elif action == 'listservicios':
                try:
                    lista = []
                    if 'idarea' in request.GET:
                        id = int(request.GET['idarea'])
                        servicios = ServicioConfigurado.objects.filter(status=True, serviciocita__departamentoservicio_id=id, mostrar=True )

                        for s in servicios:
                            text = str(s.serviciocita.nombre)
                            lista.append([encrypt(s.id), text])
                        return JsonResponse({'result': True, 'lista': lista})
                    else:
                        id = int(request.GET['id'])
                        servicios = ServicioConfigurado.objects.filter(status=True,  serviciocita__departamentoservicio_id=id, mostrar=True)
                        for r in servicios:
                            lista.append({'value':encrypt(r.id), 'text': r.serviciocita.nombre})
                        return JsonResponse({'result': True, 'data': lista})
                except Exception as e:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            elif action == 'admcalendario':
                try:
                    fecha = datetime.now().date()
                    panio = fecha.year
                    pmes = fecha.month
                    if 'mover' in request.GET:
                        mover = request.GET['mover']

                        if mover == 'anterior':
                            mes = int(request.GET['mes'])
                            anio = int(request.GET['anio'])
                            pmes = mes - 1
                            if pmes == 0:
                                pmes = 12
                                panio = anio - 1
                            else:
                                panio = anio

                        elif mover == 'proximo':
                            mes = int(request.GET['mes'])
                            anio = int(request.GET['anio'])
                            pmes = mes + 1
                            if pmes == 13:
                                pmes = 1
                                panio = anio + 1
                            else:
                                panio = anio

                    id = int(encrypt(request.GET['idservicio']))
                    idr=request.GET.get('idservicio',0)
                    s_anio = panio
                    s_mes = pmes
                    s_dia = 1
                    data['mes'] = MESES_CHOICES[s_mes - 1]
                    data['ws'] = [0, 7, 14, 21, 28, 35]
                    lista = {}
                    listahorarios = []
                    data['servicio'] = servicio = ServicioConfigurado.objects.get(pk=id)
                    # finicio = hoy - timedelta(days=servicio.numdias)
                    data['horarios'] = horarios = HorarioServicioCita.objects.filter(status=True,
                                                                                     responsableservicio__servicio_id=id,
                                                                                     # responsableservicio__responsable_id=idr,
                                                                                     mostrar=True,
                                                                                     fechafin__gte=hoy,
                                                                                     # fechainicio__gte=finicio
                                                                                     )
                    # diasreserva = servicio.numdias
                    # anio_actual = hoy.year
                    # if hoy.month != s_mes:
                    #     anioubicado=s_anio
                    #     cont=0
                    #     while anioubicado >= anio_actual:
                    #         iden = True
                    #         if anio_actual == anioubicado and cont == 0:
                    #             mes = hoy.month
                    #             mesubicado = s_mes
                    #         elif anio_actual == anioubicado:
                    #             mes = hoy.month
                    #             mesubicado = 12
                    #         elif anio_actual < anioubicado and cont == 0:
                    #             mes = 0
                    #             mesubicado = s_mes
                    #         elif anio_actual < anioubicado:
                    #             mes = 0
                    #             mesubicado = 12
                    #         while mesubicado > mes:
                    #             if anio_actual < anioubicado and iden:
                    #                 mes=1
                    #                 iden=False
                    #             numsinhorario = 0
                    #             msiguiente = calendar.monthrange(s_anio, mes)
                    #             rango = range(1, int(msiguiente[1] + 1), 1)
                    #             if mes == hoy.month:
                    #                 rango = range(int(hoy.day), int(msiguiente[1] + 1), 1)
                    #             for dia in rango:
                    #                 ban = False
                    #                 fecha = date(s_anio, mes, dia)
                    #                 numdia = fecha.weekday() + 1
                    #                 for horario in horarios:
                    #                     if horario.dia == numdia:
                    #                         ban = True
                    #                 if ban == False:
                    #                     numsinhorario += 1
                    #             diasreserva += numsinhorario
                    #             if mes == hoy.month:
                    #                 diasreserva -= msiguiente[1] - (hoy.day - 1)
                    #             else:
                    #                 diasreserva = diasreserva - msiguiente[1]
                    #
                    #             mes += 1
                    #         cont+=1
                    #         anioubicado-=1

                    for i in range(1, 43, 1):
                        dia = {i: 'no'}
                        lista.update(dia)
                    comienzo = False
                    fin = False
                    for i in lista.items():
                        try:
                            fecha = date(s_anio, s_mes, s_dia)
                            if fecha.isoweekday() == i[0] and fin is False and comienzo is False:
                                comienzo = True
                        except Exception as ex:
                            pass
                        if comienzo:
                            try:
                                fecha = date(s_anio, s_mes, s_dia)
                            except Exception as ex:
                                fin = True
                        if comienzo and fin is False:
                            dia = {i[0]: s_dia}
                            lista.update(dia)
                            listhorario = []
                            sinhorario = True
                            puedereservar = False
                            numerodia = fecha.weekday() + 1
                            if fecha >= hoy:
                                turnos = 0
                                listturnos = []
                                for horario in horarios:
                                    if horario.dia == numerodia and fecha <= horario.fechafin and fecha >= horario.fechainicio:
                                        listhorario.append(horario.id)
                                if listhorario:
                                    sinhorario = False
                                    filter = horarios.filter(status=True, pk__in=listhorario)
                                    for horario in filter:
                                        listturnos.append(horario.turno.comienza)
                                        turnos += horario.citas_disponibles(fecha)

                                monthRange = calendar.monthrange(s_anio, s_mes)
                                # if diasreserva > 0 and s_dia <= monthRange[1] and not sinhorario:
                                if s_dia <= monthRange[1] and not sinhorario:
                                    ordenadas = []
                                    ordenadas = sorted(listturnos, reverse=True)
                                    if ordenadas[0] > datetime.now().time() or fecha > datetime.now().date():
                                        puedereservar = True
                                        # diasreserva -= 1
                                diccionario = {'dia': s_dia, 'turnos': turnos, 'listahorario': listhorario,
                                               'sinhorario': sinhorario, 'fecha': fecha, 'puedereservar': puedereservar}
                                listahorarios.append(diccionario)
                            s_dia += 1

                    if horarios:
                        data['ultimafecha'] = ultimafecha = horarios.order_by('fechafin').last().fechafin
                        primerafecha = horarios.order_by('fechafin').first().fechafin
                        anio = primerafecha.year
                        rango = ultimafecha.year - primerafecha.year
                        for i in range(0, rango, 1):
                            if s_anio == anio + (i + 1):
                                anio = s_anio
                        data['year'] = anio
                    data['dias_mes'] = lista
                    data['dwnm'] = ['lunes', 'martes', 'miercoles', 'jueves', 'viernes', 'sabado', 'domingo']
                    data['dwn'] = [1, 2, 3, 4, 5, 6, 7]
                    data['daymonth'] = 1
                    data['s_anio'] = s_anio
                    data['s_mes'] = s_mes
                    data['lista'] = lista
                    data['hoy'] = hoy
                    data['hoy_dia'] = hoy.day
                    data['hoy_mes'] = hoy.month
                    data['listahorarios'] = listahorarios
                    ultimo_dia = calendar.monthrange(hoy.year, hoy.month)[1]
                    data['fechaactual'] = date(hoy.year, hoy.month, ultimo_dia)
                    data['fechacalendario'] = date(s_anio, s_mes, s_dia - 1)
                    template = get_template("adm_agendamientocitas/viewcalendario.html")
                    return JsonResponse({"result": 'ok', 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": 'bad', 'mensaje': 'Error'.format(ex)})

            elif action == 'admaddcita':
                try:
                    servicio = ServicioConfigurado.objects.get(pk=int(encrypt(request.GET['idservicio'])))
                    data['horariosid'] = horariosid = json.loads(request.GET["listaid[]"])

                    # horariosid = json.loads(request.GET["listaid[]"])
                    # if not type(eval(horariosid[0])) is int:
                    #     horariosid = list(eval(horariosid[0]))
                    if request.GET['idpersona']:
                        data['persona_ag']=Persona.objects.get(id=int(request.GET['idpersona']))
                        data['familiar'] = PersonaDatosFamiliares.objects.filter(persona__id=int(request.GET['idpersona']))
                    if request.GET['idperfil']:
                        data['perfil']=PerfilUsuario.objects.get(id=int(request.GET['idperfil']))
                    data['fecha'] = request.GET['fecha']
                    data['servicio'] = servicio
                    data['horarios'] = horario = HorarioServicioCita.objects.filter(id__in=horariosid, status=True).order_by('id')
                    data['responsables'] = horario = HorarioServicioCita.objects.filter(id__in=horariosid,
                                                                                        status=True).values_list(
                        'responsableservicio_id',
                        'responsableservicio__responsable__nombres',
                        'responsableservicio__responsable__apellido1',
                        'responsableservicio__responsable__apellido2').distinct('responsableservicio_id')
                    data['horariodia'] = HorarioServicioCita.objects.filter(id__in=horariosid, status=True).order_by(
                        'turno__comienza').first()

                    # data['horariodia'] = horario.first()
                    template = get_template("adm_agendamientocitas/modal/formadmagendar.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'mensaje': 'Error'.format(ex)})

            # Agendar cita por emergente
            elif action == 'addcitaemergente':
                try:
                    form=CitaEmergenteForm()
                    form.fields['persona'].queryset = Persona.objects.none()
                    form.fields['persona_responsable'].queryset = Persona.objects.none()
                    form.fields['personafamiliar'].queryset = Persona.objects.none()
                    form.fields['perfil'].queryset = PerfilUsuario.objects.none()
                    data['form']=form
                    data['switchery'] = True
                    template = get_template("adm_agendamientocitas/modal/formaddcitaemergente.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'mensaje': 'Error'.format(ex)})

            elif action == 'buscarpersonal':
                try:
                    resp=filtro_persona_select(request)
                    return HttpResponse(json.dumps({'status': True, 'results': resp}))
                except Exception as ex:
                    pass

            elif action == 'listresponsable':
                try:
                    lista = []
                    id = encrypt_id(request.GET['id'])
                    responsables = ResponsableServicioCita.objects.filter(status=True, servicio_id=id)
                    for r in responsables:
                        text = str(r.responsable)
                        lista.append({'value':r.responsable.id, 'text':text})
                    return JsonResponse({'result': True, 'data': lista})
                except Exception as ex:
                    return JsonResponse({"result": False, "mensaje": u"Error al obtener los datos."})

            elif action == 'listfamiliar':
                try:
                    lista = []
                    id = int(request.GET['id'])
                    familiares = PersonaDatosFamiliares.objects.filter(status=True, persona_id=id)
                    for f in familiares:
                        text = str(f.nombre)
                        lista.append({'value':f.id, 'text':text})
                    return JsonResponse({'result': True, 'data': lista})
                except Exception as ex:
                    return JsonResponse({"result": False, "mensaje": u"Error al obtener los datos."})

            elif action == 'reporte_gestorcitas':
                try:
                    rfecha = request.GET['rfechas'].split(' - ')
                    finicio, ffin = convertir_fecha_invertida(rfecha[0]), convertir_fecha_invertida(rfecha[1])

                    if request.GET['estado'] == '':
                        estadoid= [0,1,2,3,4,5,6]
                    else:
                        estadoid = request.GET['estado'].split(',')
                    if usuario.is_superuser:
                        citas_emergentes = PersonaCitaAgendada.objects.filter(
                            Q(fechacita__gte=finicio) & Q(fechacita__lte=ffin) & Q(status=True) & Q(
                                escitaemergente=True) & Q(estado__in=estadoid)).order_by('fechacita')

                        citas_planificadas = PersonaCitaAgendada.objects.filter(
                            Q(fechacita__gte=finicio) & Q(fechacita__lte=ffin) & Q(status=True) & Q(
                                escitaemergente=False) & Q(estado__in=estadoid)).order_by('fechacita')

                        subcitas = SubCitaAgendada.objects.filter(
                            Q(citaprincipal__id__in=citas_planificadas.values_list('id', flat=True)) |
                            Q(citaprincipal__id__in=citas_emergentes.values_list('id', flat=True)),
                            status=True, es_derivacion=True).order_by('fechacita')

                    else:
                        # Filtrar citas planificadas
                        citas_planificadas = PersonaCitaAgendada.objects.filter(
                            Q(fechacita__gte=finicio) & Q(fechacita__lte=ffin) & Q(status=True) &
                            Q(escitaemergente=False) & Q(persona_responsable=persona) & Q(estado__in=estadoid)).order_by('fechacita')
                        citas_emergentes = PersonaCitaAgendada.objects.filter(
                            Q(fechacita__gte=finicio) & Q(fechacita__lte=ffin) & Q(status=True) & Q(
                                escitaemergente=True) & Q(persona_responsable=persona) & Q(estado__in=estadoid)).order_by('fechacita')

                        ids_citas_pe = [x for x in citas_planificadas.values_list('id', flat=True)] + [x for x in citas_emergentes.values_list('id', flat=True)]
                    # Filtrar subcitas asociadas a las citas planificadas
                        subcitas = SubCitaAgendada.objects.filter(persona_responsable = persona, status=True, fechacita__gte=finicio, fechacita__lte=ffin,estado__in=estadoid).order_by('fechacita')
                    #     subcitas = SubCitaAgendada.objects.filter(
                    #         Q(persona_responsable=persona) &  # Responsable de la subcita
                    #         Q(status=True) &
                    #         Q(fechacita__gte=finicio) &
                    #         Q(fechacita__lte=ffin) &
                    #         Q(estado__in=estadoid) &
                    #         Q(es_derivacion=True)).distinct('citaprincipal', 'servicio').order_by('fechacita')  # Solo derivaciones
                            # Realiza el conteo de estados en las listas de citas
                    contpendientes = len(citas_emergentes.filter(estado=0)) + len(citas_planificadas.filter(estado=0)) + len(subcitas.filter(estado=0)) + len(subcitas.filter(estado=0))
                    contreservados = len(citas_emergentes.filter(estado=1)) + len(citas_planificadas.filter(estado=1)) + len(subcitas.filter(estado=1)) + len(subcitas.filter(estado=1))
                    contanulados = len(citas_emergentes.filter(estado=2)) + len(citas_planificadas.filter(estado=2)) + len(subcitas.filter(estado=2)) + len(subcitas.filter(estado=2))
                    contfinalizados = len(citas_emergentes.filter(estado=5)) + len(citas_planificadas.filter(estado=5)) + len(subcitas.filter(estado=5)) + len(subcitas.filter(estado=5))
                    contcorregir = len(citas_emergentes.filter(estado=3)) + len(citas_planificadas.filter(estado=3)) + len(subcitas.filter(estado=3)) + len(subcitas.filter(estado=3))
                    contrechazado = len(citas_emergentes.filter(estado=4)) + len(citas_planificadas.filter(estado=4)) + len(subcitas.filter(estado=4)) + len(subcitas.filter(estado=4))
                    conttramite = len(citas_emergentes.filter(estado=6)) + len(citas_planificadas.filter(estado=6)) + len(subcitas.filter(estado=6)) + len(subcitas.filter(estado=6))
                    cont_asistidas = len(citas_emergentes.filter(estado__in=[5, 6], asistio=True)) + len(
                        citas_planificadas.filter(estado__in=[5, 6], asistio=True)) + len(
                        subcitas.filter(estado__in=[5, 6], asistio=True))
                    cont_no_asistidas = len(citas_emergentes.filter(estado__in=[5, 6], asistio=False)) + len(
                        citas_planificadas.filter(estado__in=[5, 6], asistio=False)) + len(
                        subcitas.filter(estado__in=[5, 6], asistio=False))
                    total = contpendientes + contreservados + contanulados + contfinalizados + contcorregir + contrechazado + conttramite
                    qrname = 'Resumen_citas{}'.format(random.randint(1, 100000).__str__())
                # Añade los datos al contexto para pasarlo al PDF
                    return conviert_html_to_2pdf ('adm_agendamientocitas/pdfseguimientocita.html',
                                                {'pagesize': 'A4', 'data': {'hoy': datetime.now().date(),
                                                                            'fini': finicio, 'ffin': ffin,
                                                                            'persona': persona,
                                                                            'estado': estadoid,
                                                                            'citas_planificadas': citas_planificadas,
                                                                            'citas_emergentes': citas_emergentes,
                                                                            'subcitas': subcitas,
                                                                            'contpendientes': contpendientes,
                                                                            'contreservados': contreservados,
                                                                            'contanulados': contanulados,
                                                                            'contfinalizados': contfinalizados,
                                                                            'contcorregir': contcorregir,
                                                                            'contrechazado': contrechazado,
                                                                            'conttramite': conttramite,
                                                                            'total': total, 'cont_asistidas': cont_asistidas,  # Añadir total de citas asistidas
                                                                            'cont_no_asistidas': cont_no_asistidas }} ,qrname)

                except Exception as ex:
                    pass


            elif action == 'reporte_historial':
                try:

                    tipoinforme =  int(request.GET['tipoinforme'])
                    personacita = PersonaCitaAgendada.objects.get(id=int(request.GET['id']))
                    id_serviciocita = idservicio_tipoinforme(tipoinforme)
                    data['informe'] = informe = InformePsicologico.objects.filter(personacita=personacita, tipoinforme=tipoinforme, status=True).first()
                    # esto
                    if not informe:
                        raise NameError('No existe informe disponible')
                    data['detalleinforme'] = detalleinforme = DetalleInformePsicologico.objects.filter(informe=informe, status=True).order_by(
                        'estructura__orden')
                    data['procesos'] = procesos = DetalleHistorialPsicologico.objects.filter(informe=informe,status=True, proceso__mostrar=True)

                    data['proceso'] = proceso = ServicioCita.objects.get(id=id_serviciocita).procesos_tipoinforme(tipoinforme)
                    data['data'] = ''
                    data['familiares'] = familiares = informe.obtener_familiares()
                    parentescorepresentante = ''
                    if informe.personacita.espersonal:

                        result = [familiar for familiar in familiares if
                                  familiar.personafamiliar.id == informe.personacita.persona.id]
                        if result:
                            parentescorepresentante = result[0].parentesco.nombre
                        # representante = [familiar for familiar in familiares if familiar.parentesco.nombre == historial.persona.persona.nombre]

                    padres_madres = [familiar for familiar in familiares if
                                     familiar.parentesco.nombre in ['PADRE', 'MADRE']]
                    conyuge = [familiar for familiar in familiares if familiar.parentesco.nombre in ['CONYUGE']]

                    # template = 'adm_agendamientocitas/modal/informepsicopedagogico.html'

                    subtituloProcesos = {}
                    # esto
                    if informe.tipoinforme == 1:
                        template = 'adm_agendamientocitas/modal/pdfhistorialclinico.html'
                    if informe.tipoinforme == 2:
                        template = 'adm_agendamientocitas/modal/informepsicopedagogico.html'

                        for x in proceso:
                            matching_objects = [pi for pi in procesos if pi.proceso.id == x.id]
                            marcada = False
                            observacion = ''
                            if matching_objects:
                                matching_object = matching_objects[0]
                                marcada = matching_object.marcada
                                observacion = matching_object.observacion
                            detalleprocesos = []
                            detalleprocesos.append({
                                'id': x.id,
                                'descripcion': x.descripcion,
                                'marcada': marcada,
                                'observacion': observacion
                            })
                            subtitulo = matching_object.proceso.subtitulo if matching_objects else ''

                            # Verifica si el subtitulo ya existe en subtituloProcesos
                            if subtitulo in subtituloProcesos:
                                subtituloProcesos[subtitulo]['detalleprocesos'].extend(detalleprocesos.copy())
                            else:
                                # Si el subtitulo no existe, agrega una nueva entrada en subtituloProcesos
                                subtituloProcesos[subtitulo] = {
                                    'valor_subtitulo': subtitulo,
                                    'detalleprocesos': detalleprocesos.copy()
                                }
                        data['detalleprocesos'] = subtituloProcesos
                    # if informe.personacita.servicio.serviciocita.nombre == 'ÁREA PSICOMÉTRICA':
                    if informe.tipoinforme == 3:
                        template = 'adm_agendamientocitas/modal/informepsicometrico.html'
                    # if informe.personacita.servicio.serviciocita.nombre == 'REFUERZO PEDAGÓGICO':
                    if informe.tipoinforme == 4:
                        template = 'adm_agendamientocitas/modal/informe_refuerzoacademico.html'

                    qrname = 'Informe_Psicológico {}'.format(random.randint(1, 100000).__str__())
                    return conviert_html_to_2pdf(template,
                                                 {'pagesize': 'A4', 'data': {'hoy': datetime.now().date(),
                                                                             'informe': informe,
                                                                             'detalleinforme': detalleinforme,
                                                                             'procesos': procesos,
                                                                             'padremadre': padres_madres,
                                                                             'conyuge': conyuge,
                                                                             'usuario': usuario.username,
                                                                             'parentescorepresentante': parentescorepresentante,
                                                                             'detalleprocesos': subtituloProcesos
                                                                             }}, qrname)
                except Exception as ex:
                    pass


            elif action == 'reporte_refuerzoacademico':
                try:
                    cita_id = int(request.GET['id'])
                    data['cita'] = cita = PersonaCitaAgendada.objects.get(id=cita_id)

                    data['refuerzo'] = RefuerzoAcademico.objects.filter(personacita=cita, status=True).order_by('-fecha')
                    data["hoy"] = hoy
                    data["persona"] = cita.persona
                    template = 'adm_agendamientocitas/modal/pdfrefuerzoacademico.html'
                    # return JsonResponse({"result": True, 'data': template.render(data)})
                    qrname = 'Refuerzo_Academico{}'.format(random.randint(1, 100000).__str__())
                    return conviert_html_to_2pdf(template, {'pagesize': 'A4', 'data': data}, qrname)
                except Exception as ex:
                    pass
            # Registrar persona externa
            elif action == 'addpersona':
                try:
                    form = PersonaForm()
                    data['form'] = form
                    template = get_template('adm_agendamientocitas/modal/formpersona.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'consultacedulaexterno':
                try:
                    cedula = request.GET['cedula'].strip().upper()
                    pers = consultarPersona(cedula)
                    if pers:
                        if len(pers.mis_perfilesusuarios()) == 1 and pers.tiene_usuario_externo():
                            return JsonResponse({"result": True,
                                                 "apellido1": pers.apellido1,
                                                 "apellido2": pers.apellido2,
                                                 "nacimiento": pers.nacimiento.strftime('%Y-%m-%d'),
                                                 "nombres": pers.nombres,
                                                 "sexo": pers.sexo.id,
                                                 "telefono": pers.telefono,
                                                 "telefono_conv": pers.telefono_conv,
                                                 "email": pers.email})
                        else:
                            return JsonResponse({'result': 'bad', 'mensaje': 'Identificación digítada ya se encuentra registrada.'})
                    else:
                        return JsonResponse({'result': False, 'mensaje': ''})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": False, "mensaje": str(ex)})



            elif action == 'historialclinico':
                try:
                    data['title'] = u'Informe Psicológico '
                    # data['niveltitulacion'].queryset = NivelTitulacion.objects.none()
                    # data['proceso'] = proceso = cita.proceso()
                    # Obtén la instancia de PersonaCitaAgendada
                    data['cita'] = cita = PersonaCitaAgendada.objects.get(id=int(encrypt(request.GET['id'])))
                    data['tipoformulario']=tipoformulario =int(request.GET['tipoformulario'])
                    tipo_informe, id_serviciocita = tipo_informe_servicio(cita, persona)
                    if not tipo_informe:
                        raise NameError('Usted no tiene acceso a este apartado')
                    serviciopersona= cita.servicio_persona(persona)

                    asignados = asignados_id(cita)
                    data['asignados'] = asignados
                    data['cabeceraInforme'] = cabeceraInforme = cita.get_informepsicologico(tipo_informe)
                    # motivos_departamento = MotivoCita.objects.filter(departamento_asignado=asignados)
                    depservicio = cita.servicio.serviciocita.departamentoservicio.id
                    if cabeceraInforme:
                        #form = InformePsicologicoForm(initial=model_to_dict(cabeceraInforme))
                        form = InformePsicologicoForm(departamento_servicio_id=depservicio,
                                               initial=model_to_dict(cabeceraInforme))
                        data['idcab'] = cabeceraInforme.id
                    else:
                        #form = InformePsicologicoForm()
                        form = InformePsicologicoForm(departamento_servicio_id=depservicio)
                        data['idcab'] = 0
                    if serviciopersona.id != 5:
                        form.fields['grado'].widget = forms.HiddenInput()
                        form.fields['institucioneducativa'].widget = forms.HiddenInput()
                        # form.fields['archivo'].widget = forms.HiddenInput()
                    if serviciopersona.id == 7:
                        form.fields['archivo'].widget = forms.HiddenInput()
                    data['form'] = form

                    # form.fields['institucioneducativa'].queryset = InstitucionesColegio.objects.none()
                    ultimo_secuencial_procesoinforme = \
                    DetalleHistorialPsicologico.objects.filter(informe=cabeceraInforme).aggregate(Max('informe_id'))[
                        'informe_id__max']
                    procesoinforme = DetalleHistorialPsicologico.objects.filter(informe=cabeceraInforme,
                                                                                informe_id=ultimo_secuencial_procesoinforme,
                                                                                status=True)
                    data['procesoinforme'] = procesoinforme

                    ultimo_secuencial = \
                    DetalleInformePsicologico.objects.filter(informe=cabeceraInforme).aggregate(Max('informe_id'))[
                        'informe_id__max']
                    data['detalleinforme'] = detalleinforme = DetalleInformePsicologico.objects.filter(
                        informe=cabeceraInforme,
                        informe_id=ultimo_secuencial, status=True)

                    data['estructura'] = estructura = EstructuraInforme.objects.filter(servicio=serviciopersona, activo=True, tipoinforme=tipoformulario).order_by('orden')
                    detalleestructura = []
                    for x in estructura:
                        matching_objects_estructura = [pi for pi in detalleinforme if pi.estructura.id == x.id]
                        observacion = ''
                        if matching_objects_estructura:
                            matching_objects_estructura = matching_objects_estructura[0]
                            observacion = matching_objects_estructura.descripcion
                        detalleestructura.append({
                            'id': x.id,
                            'titulo': x.titulo,
                            'observacion': observacion
                        })
                        data['detalleestructura'] = detalleestructura

                    subtituloProcesos = {}

                    data['proceso'] = proceso = serviciopersona.procesos().filter(tipoinforme=tipo_informe)
                    if serviciopersona.id == 5:
                        for x in proceso:
                            matching_objects = [pi for pi in procesoinforme if pi.proceso.id == x.id]
                            marcada = False
                            observacion = ''

                            if matching_objects:
                                matching_object = matching_objects[0]
                                marcada = matching_object.marcada
                                observacion = matching_object.observacion

                            detalleprocesos = []
                            # subtitulo = matching_object.proceso.subtitulo
                            subtitulo = x.subtitulo

                            detalleprocesos.append({
                                'id': x.id,
                                'descripcion': x.descripcion,
                                'marcada': marcada,
                                'observacion': observacion,
                                'subtitulo': subtitulo
                            })

                            # Verifica si el subtitulo ya existe en subtituloProcesos

                            if subtitulo in subtituloProcesos:
                                subtituloProcesos[subtitulo]['detalleprocesos'].extend(detalleprocesos.copy())
                            else:
                                # Si el subtitulo no existe, agrega una nueva entrada en subtituloProcesos
                                subtituloProcesos[subtitulo] = {
                                    'valor_subtitulo': subtitulo,
                                    'detalleprocesos': detalleprocesos.copy()
                                }
                        data['detalleprocesos'] = subtituloProcesos

                    else:
                        detalleprocesos = []
                        for x in proceso:
                            matching_objects = [pi for pi in procesoinforme if pi.proceso.id == x.id]
                            marcada = False
                            observacion = ''
                            if matching_objects:
                                matching_object = matching_objects[0]
                                marcada = matching_object.marcada
                                observacion = matching_object.observacion
                            detalleprocesos.append({
                                'id': x.id,
                                'descripcion': x.descripcion,
                                'marcada': marcada,
                                'observacion': observacion
                            })
                        data['detalleprocesos'] = detalleprocesos

                    if not cita.espersonal:
                        # data['usuariohistorial'] = usuariohistorial = cita.persona.id
                        data['usuariohistorial'] = usuariohistorial = cita.id
                    else:
                        # data['usuariohistorial'] = usuariohistorial = cita.familiar.id
                        data['usuariohistorial'] = usuariohistorial = cita.id

                    if not persona.id in cita.responsables_subcitas_list() and not usuario.is_superuser:
                        messages.error(request, 'Usted no tiene acceso a esta solicitud.')
                        return redirect(request.path)
                    data['navactivo'] = tipoformulario
                    data['tipo_informe']=tipo_informe
                    template = 'adm_agendamientocitas/modal/formhistorialcitas.html'
                    return render(request, template, data)
                except Exception as ex:
                    print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
                    pass

            # elif action == 'buscarinstituciones':
            #     try:
            #         q = request.GET.get('q', '').upper().strip()
            #         instituciones = InstitucionesColegio.objects.filter(nombre__icontains=q)[:5]
            #
            #         lista = [{'id': inst.id, 'text': inst.nombre} for inst in instituciones]
            #
            #         return JsonResponse({'results': lista})
            #     except Exception as e:
            #         return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})


            elif action == 'familiarhistorialclinico':
                try:
                    data['title'] = u'Familiares'
                    data['cita'] = cita = PersonaCitaAgendada.objects.get(id=int(encrypt(request.GET['id'])))
                    data['asignados'] = asignados_id(cita)
                    if cita.familiar:
                        data['listado'] = listado = cita.familiar.personafamiliar.familiares().order_by('-id')
                    else:
                        data['listado'] = listado = cita.persona.familiares().order_by('-id')
                        # Filtra los historiales relacionados con la fecha de cita de la persona
                    data['citas'] = citas = InformePsicologico.objects.filter(personacita__persona_id=cita.persona.id, status=True)

                    if not cita.espersonal:
                            # data['usuariohistorial'] = usuariohistorial = cita.persona.id
                        data['usuariohistorial'] = usuariohistorial = cita.id
                    else:
                            # data['usuariohistorial'] = usuariohistorial = cita.familiar.id
                        data['usuariohistorial'] = usuariohistorial = cita.id

                    if not persona.id in cita.responsables_subcitas_list() and not usuario.is_superuser:
                        messages.error(request, 'Usted no tiene acceso a esta solicitud.')
                        return redirect(request.path)
                    data['proceso'] = proceso = cita.proceso()
                    request.session['viewactivo'] = 1
                    data['navactivo'] = 'familiares'
                    template = 'adm_agendamientocitas/modal/formfamiliares.html'
                    return render(request, template, data)

                except Exception as ex:
                    pass


            # elif action == 'verhistorial':
            #     try:
            #         data = {}
            #         detallehistoriales_data = []
            #         # data['historial']=\
            #         historial = HistorialClinicoCitas.objects.get(id=int(request.GET['id']))
            #         detallehistoriales = DetalleHistorialCitas.objects.filter(historial__id=historial.id)
            #         data['historial'] = {
            #             'id': historial.id,
            #             'motivoconsulta': historial.motivoconsulta.id,
            #             'descripcionmotivoconsulta': historial.descripcionmotivoconsulta,
            #             'desconductal': historial.desconductal,
            #             'desefectiva': historial.desefectiva,
            #             'nivelacademico': historial.nivelacademico,
            #             'diagnostico': historial.diagnostico,
            #             'conclusiones': historial.conclusiones,
            #             'recomendaciones': historial.recomendaciones,
            #             'archivo': historial.archivo.url if historial.archivo else None
            #         }
            #         for detallehistorial in detallehistoriales:
            #             detallehistorial_data = {
            #
            #                 'id_proceso': detallehistorial.proceso.id,
            #                 'marcada': detallehistorial.marcada,
            #                 'observacion': detallehistorial.observacion,
            #             }
            #             detallehistoriales_data.append(detallehistorial_data)
            #         data['historial']['detallehistoriales'] = detallehistoriales_data
            #         data['readonly'] = True
            #
            #         return JsonResponse({"result": True, 'data': data})
            #
            #     except Exception as ex:
            #         pass

            elif action == 'addfamiliar':
                try:
                    form = FamiliarForm()
                    idcita=encrypt_id(request.GET['idex'])
                    cita= PersonaCitaAgendada.objects.get(id=idcita)
                    idp=cita.get_persona().id
                    data['idp'] =idp
                    visible_fields = form.visible_fields()
                    total_fields = len(visible_fields)
                    lista = [(1, 'Informació Básica', visible_fields[:19]),
                             (2, 'Información Laboral', visible_fields[19:27]),
                             (3, 'Discapacidad', visible_fields[27:total_fields])
                             ]
                    data['form'] = lista
                    data['switchery'] = True
                    template = get_template('th_hojavida/informacionpersonal/modal/formfamiliar.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'editfamiliar':
                try:
                    data['title'] = u'Editar familiar'
                    data['id'] = id = encrypt_id(request.GET['id'])
                    idp = encrypt_id(request.GET['idex'])
                    data['familiar'] = familiar = PersonaDatosFamiliares.objects.get(pk=id)
                    cita = PersonaCitaAgendada.objects.get(id=idp)
                    data['idp'] = cita.get_persona().id
                    #data['familiar'] = familiar = PersonaDatosFamiliares.objects.get(pk=id)
                    apellido1, apellido2, nombres, editdiscapacidad, sexo = '', '', familiar.nombre, False, None
                    if familiar.personafamiliar:
                        apellido1 = familiar.personafamiliar.apellido1
                        apellido2 = familiar.personafamiliar.apellido2
                        nombres = familiar.personafamiliar.nombres
                        sexo = familiar.personafamiliar.sexo
                        perfil_f = familiar.personafamiliar.perfilinscripcion_set.filter(status=True).first()
                        estado = True if perfil_f and perfil_f.estadoarchivodiscapacidad == 2 else False
                        if len(familiar.personafamiliar.mis_perfilesusuarios()) == 1 and familiar.personafamiliar.tiene_usuario_externo():
                            editdiscapacidad = True
                        elif not estado:
                            editdiscapacidad = True
                    else:
                        editdiscapacidad = True
                    banderacedula = 0
                    if familiar.cedulaidentidad:
                        banderacedula = 1
                    data['banderacedula'] = banderacedula
                    data['edit_d'] = editdiscapacidad
                    form = FamiliarForm(initial={'identificacion': familiar.identificacion,
                                                 'parentesco': familiar.parentesco,
                                                 'nombre': nombres,
                                                 'apellido1': apellido1,
                                                 'apellido2': apellido2,
                                                 'sexo': sexo,
                                                 'nacimiento': familiar.nacimiento,
                                                 'fallecido': familiar.fallecido,
                                                 'tienediscapacidad': familiar.tienediscapacidad,
                                                 'telefono': familiar.telefono,
                                                 'niveltitulacion': familiar.niveltitulacion,
                                                 'ingresomensual': familiar.ingresomensual,
                                                 'formatrabajo': familiar.formatrabajo,
                                                 'telefono_conv': familiar.telefono_conv,
                                                 'trabajo': familiar.trabajo,
                                                 'convive': familiar.convive,
                                                 'sustentohogar': familiar.sustentohogar,
                                                 'essustituto': familiar.essustituto,
                                                 'autorizadoministerio': familiar.autorizadoministerio,
                                                 'tipodiscapacidad': familiar.tipodiscapacidad,
                                                 'porcientodiscapacidad': familiar.porcientodiscapacidad,
                                                 'carnetdiscapacidad': familiar.carnetdiscapacidad,
                                                 'institucionvalida': familiar.institucionvalida,
                                                 'tipoinstitucionlaboral': familiar.tipoinstitucionlaboral,
                                                 'tienenegocio': familiar.tienenegocio,
                                                 'esservidorpublico': familiar.esservidorpublico,
                                                 'bajocustodia': familiar.bajocustodia,
                                                 'tienenegocio': familiar.tienenegocio,
                                                 'cedulaidentidad': familiar.cedulaidentidad,
                                                 'ceduladiscapacidad': familiar.ceduladiscapacidad,
                                                 'archivoautorizado': familiar.archivoautorizado,
                                                 'cartaconsentimiento': familiar.cartaconsentimiento,
                                                 'archivocustodia': familiar.archivocustodia,
                                                 'centrocuidado': familiar.centrocuidado,
                                                 'centrocuidadodesc': familiar.centrocuidadodesc,
                                                 'negocio': familiar.negocio, })
                    visible_fields = form.visible_fields()
                    total_fields = len(visible_fields)
                    lista = [(1, 'Informació Básica', visible_fields[:19]),
                             (2, 'Información Laboral', visible_fields[19:27]),
                             (3, 'Discapacidad', visible_fields[27:total_fields])
                             ]
                    form.edit()
                    data['form'] = lista
                    data['switchery'] = True
                    template = get_template('th_hojavida/informacionpersonal/modal/formfamiliar.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass
            elif action == 'previsualizarmotivo':
                try:
                    data['id'] = citaid = encrypt_id(request.GET['id'])
                    data['cita'] = cita = PersonaCitaAgendada.objects.get(id=citaid)
                    # data['motivo'] = MotivoCita.objects.get(id=id)
                    template = get_template('adm_agendamientocitas/modal/previsualizarmotivocita.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'consultacedula':
                try:
                    cedula = request.GET['cedula'].strip().upper()
                    id=request.GET.get('id',0)
                    idp=request.GET.get('idpersona',0)
                    persona_=Persona.objects.get(id=idp)
                    datospersona = Persona.objects.filter(Q(pasaporte=cedula) | Q(cedula=cedula) | Q(pasaporte=('VS'+cedula)) | Q(cedula=cedula[2:]), status=True).first()
                    if datospersona:
                        if datospersona == persona:
                            return JsonResponse({"result": "bad", 'mensaje': 'No puede agregar su propia cédula.'})
                        if persona_.personadatosfamiliares_set.filter(identificacion=cedula, status=True).exclude(id=id).exists():
                            return JsonResponse({"result": "bad", 'mensaje':'Identificación ingresada ya se encuentra registrada en un familiar'})
                        perfil_i=datospersona.perfilinscripcion_set.filter(status=True).first()
                        editdiscapacidad=False
                        if len(datospersona.mis_perfilesusuarios()) == 1 and datospersona.tiene_usuario_externo():
                            editdiscapacidad = True
                        elif perfil_i and not perfil_i.estadoarchivodiscapacidad == 2:
                            editdiscapacidad = True
                        context={}
                        if perfil_i and perfil_i.tienediscapacidad:
                            context={'tipodiscapacidad':perfil_i.tipodiscapacidad.id if perfil_i.tipodiscapacidad else False,
                                    'tienediscapacidad':perfil_i.tienediscapacidad,
                                    'porcientodiscapacidad':perfil_i.porcientodiscapacidad,
                                    'carnetdiscapacidad':perfil_i.carnetdiscapacidad,
                                    'institucionvalida':perfil_i.institucionvalida.id if perfil_i.institucionvalida else False,
                                    'ceduladiscapacidad':perfil_i.archivo.url if perfil_i.archivo else False ,
                                    'archivoautorizado':perfil_i.archivovaloracion.url if perfil_i.archivovaloracion else False,
                                    }
                        return JsonResponse({"result": "ok",
                                             "apellido1": datospersona.apellido1,
                                             "apellido2": datospersona.apellido2,
                                             "nacimiento": datospersona.nacimiento.strftime('%Y-%m-%d'),
                                             "nombres": datospersona.nombres,
                                             "telefono": datospersona.telefono,
                                             "sexo": datospersona.sexo.id,
                                             "telefono_conv": datospersona.telefono_conv,
                                             "puedeeditar": editdiscapacidad,
                                             "perfil_i":context})
                    else:
                        return JsonResponse({"result": "no"})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": str(ex)})

            elif action == 'delfamiliar':
                try:
                    # persona = request.session['persona']
                    data['id'] = id = encrypt_id(request.GET['idex'])
                    data['familiar'] = familiar = PersonaDatosFamiliares.objects.get(pk=id)
                    # familiar = PersonaDatosFamiliares.objects.get(pk=encrypt_id(request.POST['id']))
                    familiar.status = False
                    familiar.save(request)
                    log(u'Elimino familiar: %s' % persona, request, "del")
                    res_js = {'error': False}
                except Exception as ex:
                    transaction.set_rollback(True)
                    err = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
                    msg_err = f"Ocurrio un error: {ex.__str__()}. {err}"
                    res_js = {'error': True, 'mensaje': msg_err}
                return JsonResponse(res_js)

            return HttpResponseRedirect(request.path)
        else:
            try:
                data['title'] = u'Total de Citas Agendadas'
                responsable, estadocita, perfil, desde, hasta, search, filtro, url_vars = request.GET.get('responsable',
                                                                                                          ''), \
                                                                                          request.GET.get('estadocita',
                                                                                                          ''), \
                                                                                          request.GET.get('perfil', ''), \
                                                                                          request.GET.get('desde', ''), \
                                                                                          request.GET.get('hasta', ''), \
                                                                                          request.GET.get('s', ''), Q(
                    status=True), ''
                if es_responsable and not request.user.has_perm('sga.puede_revisar_total_citas') or responsable:
                    idpersona = persona.id
                    ids_citas = []
                    if responsable:
                        data['responsable'] = idresponsable = int(encrypt(responsable))
                        url_vars += "&responsable={}".format(responsable)
                        idpersona = idresponsable
                    if es_responsable:
                        ids_citas = SubCitaAgendada.objects.filter(status=True, persona_responsable=persona). \
                            exclude(citaprincipal__persona_responsable=persona). \
                            values_list('citaprincipal_id', flat=True).order_by('citaprincipal_id').distinct()

                    filtro = filtro & Q(persona_responsable_id=idpersona) | Q(id__in=ids_citas)

                if perfil:
                    data['perfil'] = perfil = int(perfil)
                    url_vars += "&perfil={}".format(perfil)
                    if perfil == 1:
                        filtro = filtro & Q(perfil__inscripcion_id__gte=0)
                    elif perfil == 2:
                        filtro = filtro & Q(perfil__administrativo_id__gte=0)
                    elif perfil == 3:
                        filtro = filtro & Q(perfil__profesor_id__gte=0)
                    elif perfil == 4:
                        filtro = filtro & Q(perfil__externo_id__gte=0)

                if estadocita:
                    data['estadocita'] = estadocita = int(estadocita)
                    url_vars += "&estadocita={}".format(estadocita)
                    filtro = filtro & Q(estado=estadocita)

                formato = "%Y-%m-%d"
                hoy = datetime.now().date()

                if desde:
                    data['desde'] = desde
                    url_vars += "&desde={}".format(desde)
                    filtro = filtro & Q(fechacita__gte=desde)

                if hasta:
                    data['hasta'] = hasta
                    url_vars += "&hasta={}".format(hasta)
                    filtro = filtro & Q(fechacita__lte=hasta)

                desdefiltro = datetime(hoy.year, hoy.month, 1)
                hastafiltro = datetime(hoy.year, hoy.month, calendar.monthrange(hoy.year, hoy.month)[1])

                if search:
                    s = search.split(' ')
                    if len(s) == 1:
                        filtro = filtro & (
                                Q(persona__nombres__icontains=search) | Q(persona__cedula__icontains=search) | Q(
                            codigo__icontains=search))
                    if len(s) >= 2:
                        filtro = filtro & (
                                Q(persona__apellido1__icontains=s[0]) & Q(persona__apellido2__icontains=s[1]))
                    url_vars += '&s=' + search
                    data['search'] = search

                # listado = PersonaCitaAgendada.objects.filter(filtro).order_by('-id')
                listado = PersonaCitaAgendada.objects.filter(filtro).order_by('-id', '-fechacita')

                paging = MiPaginador(listado, 20)
                p = 1
                try:
                    paginasesion = 1
                    if 'paginador' in request.session:
                        paginasesion = int(request.session['paginador'])
                    if 'page' in request.GET:
                        p = int(request.GET['page'])
                    else:
                        p = paginasesion
                    try:
                        page = paging.page(p)
                    except:
                        p = 1
                    page = paging.page(p)
                except:
                    page = paging.page(p)
                request.session['paginador'] = p
                data['paging'] = paging
                data['usuario'] = usuario
                data['rangospaging'] = paging.rangos_paginado(p)
                data['page'] = page
                data["url_vars"] = url_vars
                data['listado'] = page.object_list
                data['email_domain'] = EMAIL_DOMAIN
                data['estados'] = ESTADO_SOLICITUD_SERVICIO
                data['responsableservicios'] = ResponsableServicioCita.objects.filter(status=True,
                                                                                      activo=True).order_by(
                    'responsable_id').distinct('responsable_id')
                # CONTADOR
                data['contpendientes'] = len(listado.filter(estado=0))
                data['contreservados'] = len(listado.filter(estado=1))
                data['contanulados'] = len(listado.filter(estado=2))
                data['contfinalizados'] = len(listado.filter(estado=5))
                data['contcorregir'] = len(listado.filter(estado=3))
                data['contrechazado'] = len(listado.filter(estado=4))
                data['conttramite'] = len(listado.filter(estado=6))
                data['total'] = len(listado)

                data['desdefiltro'] = desdefiltro
                data['hastafiltro'] = hastafiltro
                # GENERAR REPORTE DE CITAS

                if 'exportar_excel' in request.GET:
                    wb = openxl.Workbook()
                    wb["Sheet"].title = "Reporte_Citas"
                    ws = wb.active
                    style_title = openxlFont(name='Arial', size=16, bold=True)
                    style_cab = openxlFont(name='Arial', size=10, bold=True)
                    alinear = alin(horizontal="center", vertical="center")
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=Reporte_de_citas_' + str(
                        random.randint(1, 10000)) + '.xlsx'

                    # Define column widths
                    column_widths = {
                        'A': 5, 'B': 25, 'C': 15, 'D': 10, 'E': 25, 'F': 20, 'G': 20, 'H': 20, 'I': 15, 'J': 25
                    }
                    for col, width in column_widths.items():
                        ws.column_dimensions[col].width = width

                    # Merge and style title cell
                    ws.merge_cells('A1:J1')
                    ws['A1'] = 'REPORTE DE CITAS'
                    celda1 = ws['A1']
                    celda1.font = style_title
                    celda1.alignment = alinear

                    # Define column headers
                    columns = [
                        "N°", "NOMBRES Y APELLIDOS", "CÉDULA", "EDAD", "CARRERA", "MODALIDAD",
                        "TIPO DE SERVICIO", "FECHA DE CONSULTA", "ESTADO DE CONSULTA", "RESPONSABLE"
                    ]
                    row_num = 3
                    for col_num, column_title in enumerate(columns, start=1):
                        celda = ws.cell(row=row_num, column=col_num, value=column_title)
                        celda.font = style_cab
                        celda.alignment = alinear

                    mensaje = 'NO REGISTRA'
                    # Populate data
                    row_num = 4
                    #listadoFiltrado = listado.filter(persona_responsable=persona)
                    for numero, cita in enumerate(listado, start=1):
                        # Correctly calling the edad method from the persona object
                        edad = cita.persona.edad()
                        carrera = cita.perfil.inscripcion.carrera if cita.perfil.es_estudiante() else 'NO REGISTRA'
                        modalidad = cita.perfil.inscripcion.modalidad if cita.perfil.es_estudiante() else 'USUARIO EXTERNO'
                        # tipo_servicio = cita.servicio.get_tipo_display()  # Assumes a method get_tipo_display() exists
                        estado = cita.get_estado_display()

                        ws.cell(row=row_num, column=1, value=numero)
                        ws.cell(row=row_num, column=2, value=str(cita.persona))  # Ensure this is a string
                        ws.cell(row=row_num, column=3, value=str(cita.persona.cedula))  # Ensure this is a string
                        ws.cell(row=row_num, column=4, value=str(edad))  # Ensure this is a string or number
                        ws.cell(row=row_num, column=5, value=str(carrera))  # Ensure this is a string
                        ws.cell(row=row_num, column=6, value=str(modalidad))  # Ensure this is a string
                        ws.cell(row=row_num, column=7, value=str(cita.servicio.serviciocita))  # Ensure this is a string
                        ws.cell(row=row_num, column=8, value=str(cita.fechacita))  # Ensure this is a string or number
                        ws.cell(row=row_num, column=9, value=str(estado))  # Ensure this is a string
                        ws.cell(row=row_num, column=10, value=str(cita.persona_responsable))  # Ensure this is a string
                        row_num += 1

                    wb.save(response)
                    return response
                return render(request, 'adm_agendamientocitas/viewcitas.html', data)
            except Exception as ex:
                print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
                print(str(ex))
                return HttpResponseRedirect(f'/?info={str(ex)}')


def asignados_id(cita):
    subcitas = cita.subcita_set.filter(status=True).order_by('servicio__serviciocita_id').distinct()
    lista=[]
    lista.append(cita.servicio.serviciocita.id)
    for s in subcitas:
        lista.append(s.servicio.serviciocita.id)
    return lista


def tipo_informe_servicio(cita, persona):
    # lista= []
    tiposervicios={5:2,6:1,7:3,8:4}
    subcita = cita.subcitas().filter(es_derivacion=True,persona_responsable=persona).exclude(citaprincipal__persona_responsable=persona).order_by('servicio_id').distinct().first()
    # for c in subcitas:
    #     lista.append(tiposervicios[c.servicio.id])
    if subcita:
        return tiposervicios[subcita.servicio.serviciocita.id], subcita.servicio.serviciocita.id
    elif cita.servicio.serviciocita.id in [5,6,7,8]:
        return tiposervicios[cita.servicio.serviciocita.id], cita.servicio.serviciocita.id
    else:
        return None, None

def idservicio_tipoinforme(tipoinforme):
    # lista= []
    idservicicita={2:5,1:6,3:7,4:8}
    return idservicicita[tipoinforme]
