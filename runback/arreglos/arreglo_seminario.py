#!/usr/bin/env python
# -*- coding: utf-8 -*-
import os
import sys
import openpyxl
# import urllib2

# Full path and name to your csv file
import unicodedata
# from django.db.backends.oracle.base import to_unicode
# from apt.package import Record

import xlrd
# from __builtin__ import file
# from IPython.lib.editorhooks import mate
from django.http import HttpResponse
# from numpy.core.records import record
# from numpy.matrixlib.defmatrix import matrix
from setuptools.windows_support import hide_file
from urllib3 import request
from docx import Document

from settings import EMAIL_INSTITUCIONAL_AUTOMATICO, EMAIL_DOMAIN, PROFESORES_GROUP_ID, \
    RESPONSABLE_BIENES_ID, ALUMNOS_GROUP_ID, USA_TIPOS_INSCRIPCIONES, TIPO_INSCRIPCION_INICIAL, DIAS_MATRICULA_EXPIRA, \
    CLAVE_USUARIO_CEDULA, CHEQUEAR_CONFLICTO_HORARIO

SITE_ROOT = os.path.dirname(os.path.realpath(__file__))


# csv_filepathname = "cbaja.xlsx"
# csv_filepathname2 = "parroquia.txt"
# csv_filepathname2 = "notas2015salida.csv"
# csv_filepathname = "notas_ingles_150515.csv"
# csv_filepathname = "tesis.csv"
# csv_filepathname = "registro_libros.csv"
# csv_filepathname = "cuentas.csv"
# csv_filepathname = "deudas.csv"
# csv_filepathname = "records.csv"
# csv_filepathname2 = "deudas.csv"G
#csv_filepathname2 = "notas.csv"
# csv_filepathname= "MATERIAS.txt"
# csv_filepathname2 = "historico.csv"
# csv_filepathname2 = "fallos_notas_ingles.csv"
csv_filepathname2 = "egresadofalta.csv"

# your_djangoproject_home=os.path.split(SITE_ROOT)[0]
#
# sys.path.append(your_djangoproject_home)
# os.environ['DJANGO_SETTINGS_MODULE'] = 'settings'



# SITE_ROOT = os.path.dirname(os.path.realpath(__file__))
YOUR_PATH = os.path.dirname(os.path.realpath(__file__))
# print(f"YOUR_PATH: {YOUR_PATH}")
SITE_ROOT = os.path.dirname(os.path.dirname(YOUR_PATH))
SITE_ROOT = os.path.join(SITE_ROOT, '')
# print(f"SITE_ROOT: {SITE_ROOT}")
your_djangoproject_home = os.path.split(SITE_ROOT)[0]
# print(f"your_djangoproject_home: {your_djangoproject_home}")
sys.path.append(your_djangoproject_home)

from django.core.wsgi import get_wsgi_application
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "settings")
application = get_wsgi_application()




# def download(url, NOMBRE):
#     try:
#         furl = urllib2.urlopen(url)
#         f = file("%s.png"%NOMBRE,'wb')
#         f.write(furl.read())
#         f.close()
#     except:
#         print('Unable to download file')

def copia_clases(cla, inicio):
    if LeccionGrupo.objects.filter(profesor=cla.materia.profesor_principal(), turno=cla.turno, fecha=inicio).exists():
        lecciongrupo = LeccionGrupo.objects.get(profesor=cla.materia.profesor_principal(), turno=cla.turno, fecha=inicio)
    else:
        lecciongrupo = LeccionGrupo(profesor=cla.materia.profesor_principal(),
                                    turno=cla.turno,
                                    aula=cla.aula,
                                    dia=cla.dia,
                                    fecha=inicio,
                                    horaentrada=cla.turno.comienza,
                                    horasalida=cla.turno.termina,
                                    abierta=False,
                                    automatica=True,
                                    contenido='',
                                    estrategiasmetodologicas='',
                                    observaciones='')
        lecciongrupo.save()
    leccion = Leccion(clase=cla,
                      fecha=inicio,
                      horaentrada=cla.turno.comienza,
                      abierta=False,
                      horasalida=cla.turno.termina,
                      contenido='',
                      estrategiasmetodologicas='',
                      observaciones='')
    leccion.save()
    lecciongrupo.lecciones.add(leccion)
    asignadas = materia.materiaasignada_set.filter(matricula__estado_matricula__in=[2,3])
    for materiaasignada in asignadas:
        asistencialeccion = AsistenciaLeccion(leccion=leccion,
                                              materiaasignada=materiaasignada,
                                              asistio=True)
        asistencialeccion.save()
        materiaasignada.save(actualiza=True)
        materiaasignada.actualiza_estado()
    lecciongrupo.save()


def to_unicode(s):
    if isinstance(s, unicode):
        return s

    from locale import getpreferredencoding
    for cp in (getpreferredencoding(), "cp1255", "cp1250"):
        try:
            return unicode(s, cp)
        except UnicodeDecodeError:
            pass
    raise Exception("Conversion to unicode failed")

def fechatope(fecha):
    contador = 0
    nuevafecha = fecha
    while contador < DIAS_MATRICULA_EXPIRA:
        nuevafecha = nuevafecha + timedelta(1)
        if nuevafecha.weekday() != 5 and nuevafecha.weekday() != 6:
            contador += 1
    return nuevafecha

def calculate_username(persona, variant=1):
    alfabeto = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n', 'o', 'p', 'q', 'r', 's', 't', 'u', 'v', 'w', 'x', 'y', 'z', '0', '1', '2', '3', '4', '5', '6', '7', '8', '9']
    s = persona.nombres.lower().split(' ')
    while '' in s:
        s.remove('')
    if persona.apellido2:
        usernamevariant = s[0][0] + persona.apellido1.lower() + persona.apellido2.lower()[0]
    else:
        usernamevariant = s[0][0] + persona.apellido1.lower()
    usernamevariant = usernamevariant.replace(' ', '').replace(u'ñ', 'n').replace(u'á', 'a').replace(u'é', 'e').replace(u'í', 'i').replace(u'ó', 'o').replace(u'ú', 'u')
    usernamevariantfinal = ''
    for letra in usernamevariant:
        if letra in alfabeto:
            usernamevariantfinal += letra
    if variant > 1:
        usernamevariantfinal += str(variant)
    if not User.objects.filter(username=usernamevariantfinal).exclude(persona=persona).exists():
        return usernamevariantfinal
    else:
        return calculate_username(persona, variant + 1)

def generar_usuario(persona, usuario, group_id):
    password = DEFAULT_PASSWORD
    if CLAVE_USUARIO_CEDULA:
        password = persona.cedula
    user = User.objects.create_user(usuario, '', password)
    user.save()
    persona.usuario = user
    persona.save()
    persona.cambiar_clave()
    g = Group.objects.get(pk=group_id)
    g.user_set.add(user)
    g.save()

from sga.models import *
from sagest.models import *

# from bib.models import *
# from med.models import *
from datetime import date
# from sga.docentes import calculate_username
import csv

dataWriter = csv.writer(open(csv_filepathname2,'ab'), delimiter=';')
# dataReader = csv.reader(open(csv_filepathname,"rU"), delimiter=';')
# dataReader2 = csv.reader(open(csv_filepathname2,"rU"), delimiter=';')


# import xlrd

# workbook = xlrd.open_workbook("productos_unemi.xlsx")
# sheet = workbook.sheet_by_index(0)

# workbook = xlrd.open_workbook("producto.xlsx")
# sheet = workbook.sheet_by_index(0)

def convertirfecha(fecha):
    try:
        return date(int(fecha[6:10]),int(fecha[3:5]),int(fecha[0:2]))
    except Exception as ex:
        return datetime.now().date()

def convertirfechahora(fecha):
    try:
        return datetime(int(fecha[0:4]), int(fecha[5:7]), int(fecha[8:10]),int(fecha[11:13]),int(fecha[14:16]),int(fecha[17:19]))
    except Exception as ex:
        return datetime.now()


def convertirfecha2(fecha):
    try:
        return date(int(fecha[0:4]),int(fecha[5:7]),int(fecha[8:10]))
    except Exception as ex:
        return datetime.now().date()

def remover_caracteres_especiales(cadena):
    s = ''.join((c for c in unicodedata.normalize('NFD',unicode(cadena)) if unicodedata.category(c) != 'Mn'))
    return s.decode()

def fecha_letra(valor):
    Mes = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    d = int(valor[0:2])
    m = int(valor[3:5])
    a = int(valor[6:10])
    if d == 1:
        return  u"al %s día del mes de %s del %s" % (numero_a_letras(d),str(Mes[m - 1]),numero_a_letras(a))
    else:
        return u"a los %s días del mes de %s del %s" % (numero_a_letras(d),str(Mes[m - 1]),numero_a_letras(a))

def daterange(start_date, end_date):
    for n in range(int((end_date - start_date).days)):
        yield start_date + timedelta(n)

MONEDA_SINGULAR = 'dolar'
MONEDA_PLURAL = 'dolares'

CENTIMOS_SINGULAR = 'centavo'
CENTIMOS_PLURAL = 'centavos'

MAX_NUMERO = 999999999999

UNIDADES = (
    'cero',
    'uno',
    'dos',
    'tres',
    'cuatro',
    'cinco',
    'seis',
    'siete',
    'ocho',
    'nueve'
)

DECENAS = (
    'diez',
    'once',
    'doce',
    'trece',
    'catorce',
    'quince',
    'dieciseis',
    'diecisiete',
    'dieciocho',
    'diecinueve'
)

DIEZ_DIEZ = (
    'cero',
    'diez',
    'veinte',
    'treinta',
    'cuarenta',
    'cincuenta',
    'sesenta',
    'setenta',
    'ochenta',
    'noventa'
)

CIENTOS = (
    '_',
    'ciento',
    'doscientos',
    'trescientos',
    'cuatroscientos',
    'quinientos',
    'seiscientos',
    'setecientos',
    'ochocientos',
    'novecientos'
)

def numero_a_letras(numero):
    numero_entero = int(numero)
    if numero_entero > MAX_NUMERO:
        raise OverflowError('Número demasiado alto')
    if numero_entero < 0:
        return 'menos %s' % numero_a_letras(abs(numero))
    letras_decimal = ''
    parte_decimal = int(round((abs(numero) - abs(numero_entero)) * 100))
    if parte_decimal > 9:
        letras_decimal = 'punto %s' % numero_a_letras(parte_decimal)
    elif parte_decimal > 0:
        letras_decimal = 'punto cero %s' % numero_a_letras(parte_decimal)
    if (numero_entero <= 99):
        resultado = leer_decenas(numero_entero)
    elif (numero_entero <= 999):
        resultado = leer_centenas(numero_entero)
    elif (numero_entero <= 999999):
        resultado = leer_miles(numero_entero)
    elif (numero_entero <= 999999999):
        resultado = leer_millones(numero_entero)
    else:
        resultado = leer_millardos(numero_entero)
    resultado = resultado.replace('uno mil', 'un mil')
    resultado = resultado.strip()
    resultado = resultado.replace(' _ ', ' ')
    resultado = resultado.replace('  ', ' ')
    if parte_decimal > 0:
        resultado = '%s %s' % (resultado, letras_decimal)
    return resultado

def numero_a_moneda(numero):
    if '.' in numero:
        posicion=numero.split('.')
        numero_entero = int(posicion[0])
        parte_decimal=0
        if posicion[1] != '':
            parte_decimal = int(posicion[1])
    else:
        numero_entero = int(numero)
        parte_decimal = 0
    centimos = ''
    if parte_decimal == 1:
        centimos = CENTIMOS_SINGULAR
    else:
        centimos = CENTIMOS_PLURAL
    moneda = ''
    if numero_entero == 1:
        moneda = MONEDA_SINGULAR
    else:
        moneda = MONEDA_PLURAL
    letras = numero_a_letras(numero_entero)
    letras = letras.replace('uno', 'un')
    letras_decimal = u'%s/100 DOLARES DE LOS ESTADOS UNIDOS DE NORTE AMÉRICA' % (str(parte_decimal))
    letras = u'%s %s' % (letras, letras_decimal)
    return letras

def leer_decenas(numero):
    if numero < 10:
        return UNIDADES[numero]
    decena, unidad = divmod(numero, 10)
    if unidad == 0:
        resultado = DIEZ_DIEZ[decena]
    else:
        if numero <= 19:
            resultado = DECENAS[unidad]
        elif numero <= 29:
            resultado = 'veinti%s' % UNIDADES[unidad]
        else:
            resultado = DIEZ_DIEZ[decena]
            if unidad > 0:
                resultado = '%s y %s' % (resultado, UNIDADES[unidad])
    return resultado

def leer_centenas(numero):
    centena, decena = divmod(numero, 100)
    if numero == 0:
        resultado = 'cien'
    else:
        resultado = CIENTOS[centena]
        if decena > 0:
            resultado = '%s %s' % (resultado, leer_decenas(decena))
    return resultado

def leer_miles(numero):
    millar, centena = divmod(numero, 1000)
    resultado = ''
    if (millar == 1):
        resultado = ''
    if (millar >= 2) and (millar <= 9):
        resultado = UNIDADES[millar]
    elif (millar >= 10) and (millar <= 99):
        resultado = leer_decenas(millar)
    elif (millar >= 100) and (millar <= 999):
        resultado = leer_centenas(millar)
    resultado = '%s mil' % resultado
    if centena > 0:
        resultado = '%s %s' % (resultado, leer_centenas(centena))
    return resultado

def leer_millones(numero):
    millon, millar = divmod(numero, 1000000)
    resultado = ''
    if (millon == 1):
        resultado = ' un millon '
    if (millon >= 2) and (millon <= 9):
        resultado = UNIDADES[millon]
    elif (millon >= 10) and (millon <= 99):
        resultado = leer_decenas(millon)
    elif (millon >= 100) and (millon <= 999):
        resultado = leer_centenas(millon)
    if millon > 1:
        resultado = '%s millones' % resultado
    if (millar > 0) and (millar <= 999):
        resultado = '%s %s' % (resultado, leer_centenas(millar))
    elif (millar >= 1000) and (millar <= 999999):
        resultado = '%s %s' % (resultado, leer_miles(millar))
    return resultado

def leer_millardos(numero):
    millardo, millon = divmod(numero, 1000000)
    return '%s millones %s' % (leer_miles(millardo), leer_millones(millon))

# workbook = xlrd.open_workbook("cuentacontable1.xlsx")
# sheet = workbook.sheet_by_index(0)
# linea = 1
# for rowx in range(sheet.nrows):
#     cols = sheet.row_values(rowx)
#     print linea
#     if linea > 1:
#         cuenta1 = str(cols[0])
#         if cuenta1[len(str(cols[0]))-2:] == '.0':
#             cuenta1 = str(cols[0])[:len(str(cols[0]))-2]
#         else:
#             cuenta1 = str(cols[0])
#
#         if CuentaContable.objects.filter(cuenta=cuenta1).exists():
#            cuenta =  CuentaContable.objects.filter(cuenta=cuenta1)[0]
#            cuenta.descripcion = cols[1]
#            cuenta.save(usuario=User.objects.get(pk=1))
#         else:
#             cuenta = CuentaContable(cuenta=cuenta1,
#                                     descripcion=cols[1])
#             cuenta.save(usuario=User.objects.get(pk=1))
#     linea += 1
# print 'listo cuenta'


# workbook = xlrd.open_workbook("cuentacontable.xlsx")
# sheet = workbook.sheet_by_index(0)
# linea = 1
# for rowx in range(sheet.nrows):
#     cols = sheet.row_values(rowx)
#     print linea
#     if linea > 1:
#         cuenta1 = str(cols[0])
#         if cuenta1[len(str(cuenta1))-2:] == '.0':
#             cuenta1 = str(cuenta1)[:len(str(cuenta1))-2]
#         else:
#             cuenta1 = str(cols[0])
#         cuenta =  CuentaContable.objects.filter(cuenta=cuenta1)[0]
#
#         codigopartida = str(cols[1])
#         if codigopartida[len(str(codigopartida))-2:] == '.0':
#             codigopartida = str(codigopartida)[:len(str(codigopartida))-2]
#         else:
#             codigopartida = str(cols[1])
#         partida = None
#         if Partida.objects.filter(codigo=codigopartida).exists():
#             partida = Partida.objects.filter(codigo=codigopartida)[0]
#
#         cuenta.partida=partida
#         cuenta.save(usuario=User.objects.get(pk=1))
#
#     linea += 1
# print 'listo cuenta partida'


#
#     if not TipoProducto.objects.filter(nombre=cols[1].upper()).exists():
#         tipo = TipoProducto(nombre=cols[1].upper())
#         tipo.save(usuario=User.objects.get(pk=1))
#     else:
#         tipo = TipoProducto.objects.filter(nombre=cols[1].upper())[0]
#     if not Producto.objects.filter(codigo=int(cols[2]), cuenta = cuenta).exists():
#         producto = Producto(codigo=int(cols[2]),
#                             descripcion = cols[3],
#                             unidadmedida_id = 1,
#                             tipoproducto = tipo,
#                             cuenta = cuenta,
#                             alias = '',
#                             codigobarra = '',
#                             minimo = 0,
#                             maximo = 0)
#         producto.save(usuario=User.objects.get(pk=1))
#     linea += 1

# linea = 1
# ingresoprod = IngresoProducto(proveedor_id=617,
#                               tipodocumento_id=1,
#                               numerodocumento='000-000-0000000',
#                               fechadocumento=datetime.now().date(),
#                               ordencompra='',
#                               solicitudcompra='',
#                               descripcion='MIGRACION',
#                               fecha=datetime.now().date(),
#                               fechaoperacion=datetime.now(),
#                               subtotal_base12=0,
#                               subtotal_base0=0,
#                               total_descuento=0,
#                               total_iva=0,
#                               total=0,
#                               transporte=0)
# ingresoprod.save(usuario=User.objects.get(pk=1))
# for rowx in range(sheet.nrows):
#     cols = sheet.row_values(rowx)
#     print linea
#     cuenta =  CuentaContable.objects.filter(cuenta=cols[0].strip())[0]
#     producto = Producto.objects.filter(codigo=int(cols[2]), cuenta = cuenta)[0]
#     if float(cols[4]) > 0:
#         # ITEMS
#         detalleingprod = DetalleIngresoProducto(producto=producto,
#                                                 cantidad=float(cols[4]),
#                                                 costo=Decimal(cols[5]) / float(cols[4]), 4),
#                                                 subtotal=float(cols[5]),
#                                                 descuento=0,
#                                                 coniva=False,
#                                                 valoriva=0,
#                                                 total=float(cols[5]),
#                                                 estado_id=1)
#         detalleingprod.save(usuario=User.objects.get(pk=1))
#         ingresoprod.productos.add(detalleingprod)
#         # ACTUALIZAR INVENTARIO REAL
#         if InventarioReal.objects.filter(producto=producto).exists():
#             inventarioreal = InventarioReal.objects.filter(producto=producto)[0]
#             inventarioreal.cantidad += detalleingprod.cantidad
#             inventarioreal.valor += detalleingprod.total
#             inventarioreal.costo = round(inventarioreal.valor / inventarioreal.cantidad, 2)
#             inventarioreal.save(usuario=User.objects.get(pk=1))
#         else:
#             inventarioreal = InventarioReal(producto=detalleingprod.producto,
#                                             cantidad=detalleingprod.cantidad,
#                                             costo=detalleingprod.costo,
#                                             valor=detalleingprod.total)
#             inventarioreal.save(usuario=User.objects.get(pk=1))
#         # ACTUALIZAR KARDEX
#         kardex = KardexInventario(inventario=inventarioreal,
#                                   tipomovimiento=1,
#                                   fecha=datetime.now().date(),
#                                   compra=ingresoprod,
#                                   cantidad=detalleingprod.cantidad,
#                                   costo=inventarioreal.costo,
#                                   saldo=inventarioreal.cantidad)
#         kardex.save(usuario=User.objects.get(pk=1))
#     linea += 1
# ingresoprod.save(usuario=User.objects.get(pk=1))
#




# # administrativos y departamentos
# linea = 1
# for rowx in range(sheet.nrows):
#     cols = sheet.row_values(rowx)
#     if not Persona.objects.filter(cedula__icontains=cols[0]).exists():
#         persona = Persona(nombres = cols[3],
#                           apellido1 = cols[1],
#                           apellido2 = cols[2],
#                           cedula = cols[0],
#                           pasaporte = cols[0] if cols[0][:2] == "VS" else '',
#                           nacimiento = datetime.now().date(),
#                           sexo_id = int(cols[4]),
#                           telefono = str(cols[9]),
#                           telefono_conv = "",
#                           direccion = cols[7][:99],
#                           email = "",
#                           emailinst = cols[8])
#         persona.save()
#         persona.cambiar_clave()
#         username = calculate_username(persona)
#         password = DEFAULT_PASSWORD
#         user = User.objects.create_user(username, persona.email, password)
#         user.save()
#         gru = Group.objects.get(pk=ADMINISTRATIVOS_GROUP_ID)
#         gru.user_set.add(user)
#         gru.save()
#         persona.usuario = user
#         persona.save()
#         administrativo = Administrativo(persona = persona,
#                                         contrato = '',
#                                         fechaingreso = xlrd.xldate.xldate_as_datetime(cols[6], workbook.datemode).date())
#         administrativo.save()
#         persona.crear_perfil(administrativo=administrativo)
#         print 'Creo persona'
#     else:
#         persona = Persona.objects.filter(cedula__icontains=cols[0])[0]
#         gru = Group.objects.get(pk=ADMINISTRATIVOS_GROUP_ID)
#         gru.user_set.add(persona.usuario)
#         gru.save()
#         if not persona.perfilusuario_set.filter(administrativo__isnull=False):
#             administrativo = Administrativo(persona = persona,
#                                         contrato = '',
#                                         fechaingreso = xlrd.xldate.xldate_as_datetime(cols[6], workbook.datemode).date())
#             administrativo.save()
#             persona.crear_perfil(administrativo=administrativo)
#             print 'Creo perfil'
#         else:
#             print 'Esta correcto'
#
#     if not Departamento.objects.filter(nombre=cols[5]).exists():
#         dep = Departamento(nombre=cols[5])
#         dep.save(usuario=User.objects.get(pk=1))
#     else:
#         dep = Departamento.objects.filter(nombre=cols[5])[0]
#     dep.integrantes.add(persona)
#     print linea
#     linea += 1


# # proveedores
# linea = 1
# for rowx in range(sheet.nrows):
#     cols = sheet.row_values(rowx)
#     if not Proveedor.objects.filter(identificacion=cols[0]).exists():
#         proveedor = Proveedor(nombre = cols[1],
#                               identificacion=cols[0],
#                               alias=cols[1][:10],
#                               direccion=cols[2])
#         proveedor.save(usuario=User.objects.get(pk=1))
#         print 'Creo proveedor'
#     else:
#         proveedor = Proveedor.objects.filter(identificacion=cols[0])[0]
#         proveedor.alias = proveedor.nombre.split(' ')[0]
#         proveedor.save(usuario=User.objects.get(pk=1))
#         print 'Ya existe'
#     print linea
#     linea += 1



# p = Periodo.objects.get(pk=28)
# nivel = Nivel.objects.get(pk=2)
# for r in RecordAcademico.objects.filter(asignatura__id=554):
#     if Matricula.objects.filter(inscripcion=r.inscripcion, nivel__periodo=p).exists():
#         m = Matricula.objects.filter(inscripcion=r.inscripcion, nivel__periodo=p)[0]
#         if Materia.objects.filter(asignatura_id=554, nivel__periodo=p, modulomalla__malla__carrera=r.inscripcion.carrera).exists():
#             ma = Materia.objects.filter(asignatura_id=554, nivel__periodo=p, modulomalla__malla__carrera=r.inscripcion.carrera)[0]
#         else:
#             modulomalla = r.inscripcion.malla_inscripcion().malla.modulomalla_set.filter(asignatura__id=554)[0]
#             ma = Materia(nivel_id=2,
#                          asignatura_id = 554,
#                          modulomalla=modulomalla,
#                          identificacion='',
#                          alias='',
#                          horassemanales=0,
#                          horas=modulomalla.horas,
#                          creditos=modulomalla.creditos,
#                          inicio=nivel.inicio,
#                          fin=nivel.fin,
#                          fechafinasistencias=nivel.fin,
#                          rectora=True,
#                          cerrado=True,
#                          fechacierre = nivel.fin,
#                          modeloevaluativo_id = 1,
#                          validacreditos = True,
#                          validapromedio = True,
#                          sinasistencia = True)
#             ma.save()
#         if not m.materiaasignada_set.filter(materia=ma).exists():
#             maa = MateriaAsignada(matricula = m,
#                                   materia = ma,
#                                   notafinal = r.nota,
#                                   asistenciafinal = r.asistencia,
#                                   cerrado = True,
#                                   fechacierre = p.fin,
#                                   matriculas = 1,
#                                   observaciones = 'MIGRACION',
#                                   estado_id = 1 if r.aprobada else 2,
#                                   sinasistencia = True,
#                                   fechaasignacion = ma.inicio)
#             maa.save()
#         else:
#             maa = m.materiaasignada_set.filter(materia=ma)[0]
#         maa.notafinal = r.nota
#         maa.save()
#         evalua = maa.evaluacion()
#         campo1 = maa.campo('EV.SIST.1')
#         campo1.valor = r.nota
#         campo1.save()
#         campo2 = maa.campo('EXAMEN.1')
#         campo2.valor = r.nota
#         campo2.save()
#         campo5 = maa.campo('PARC.1')
#         campo5.valor = round(campo1.valor * 0.3, 2) + round(campo2.valor * 0.2, 2)
#         campo5.save()
#         campo3 = maa.campo('EV.SIST.2')
#         campo3.valor = r.nota
#         campo3.save()
#         campo4 = maa.campo('EXAM.F.')
#         campo4.valor = r.nota
#         campo4.save()
#         campo6 = maa.campo('PARC.2')
#         campo6.valor = round(campo4.valor * 0.3, 2) + round(campo3.valor * 0.2, 2)
#         campo6.save()
#         print m.id
#     else:
#         print 'sin matricula'

# for m in Materia.objects.filter(nivel__periodo__id=28):
#     m.cupo = m.materiaasignada_set.count()
#     m.save()
#     print m.id
#     for maa in m.materiaasignada_set.all():
#         campo1 = maa.campo('EV.SIST.1')
#         campo1.valor = maa.notafinal
#         campo1.save()
#         campo2 = maa.campo('EXAMEN.1')
#         campo2.valor = maa.notafinal
#         campo2.save()
#         campo5 = maa.campo('PARC.1')
#         campo5.valor = round(campo1.valor * 0.3, 2) + round(campo2.valor * 0.2, 2)
#         campo5.save()
#         campo3 = maa.campo('EV.SIST.2')
#         campo3.valor = maa.notafinal
#         campo3.save()
#         campo4 = maa.campo('EXAM.F.')
#         campo4.valor = maa.notafinal
#         campo4.save()
#         campo6 = maa.campo('PARC.2')
#         campo6.valor = round(campo4.valor * 0.3, 2) + round(campo3.valor * 0.2, 2)
#         campo6.save()
#         print maa.id

# actualizacion datos de hoja de vida
# workbook = xlrd.open_workbook(csv_filepathname)
# sheet = workbook.sheet_by_index(0)
# n = 1
# for rowx in range(sheet.nrows):
#     print n
#     if n > 1:
#         row = sheet.row_values(rowx)
#         if n == 118:
#             pass
#         if Persona.objects.filter(cedula=row[0].strip()).exists():
#             persona = Persona.objects.filter(cedula=row[0].strip())[0]
#             persona.nacimiento = convertirfecha2(row[1].strip())
#             persona.direccion = row[34].strip()
#             persona.telefonoextension = row[4].strip()
#             persona.telefono_conv = row[5].strip()
#             persona.telefono = row[6].strip()
#             datosextension = persona.datos_extension()
#             try:
#                 if row[7].strip() == '<C>':
#                     civil = 2
#                 elif row[7].strip() == '<D>':
#                     civil = 3
#                 elif row[7].strip() == '<V>':
#                     civil = 4
#                 else:
#                     civil = 1
#             except:
#                 civil = 1
#             datosextension.estadocivil_id = civil
#             datosextension.carnetiess = row[11].strip()
#             datosextension.save()
#             # cuentabancaria
#             if row[10].strip() != '' and row[10].strip() != '0':
#                 if not CuentaBancariaPersona.objects.filter(persona=persona, banco_id=int(row[10].strip()),tipocuentabanco_id=int(row[9].strip()), numero=row[8].strip()):
#                     banco1 = Banco.objects.filter(pk=int(row[10].strip()))[0]
#                     tipocuenta1 = TipoCuentaBanco.objects.filter(pk=int(row[9].strip()))[0]
#                     cuentabanco = CuentaBancariaPersona(persona = persona,
#                                                         banco = banco1,
#                                                         tipocuentabanco = tipocuenta1,
#                                                         numero = row[8].strip(),
#                                                         verificado = True)
#                     cuentabanco.save()
#
#             personafisico = persona.datos_examen_fisico()
#             personafisico.peso = float(row[13].strip().replace(',','.'))
#             personafisico.talla = float(row[12].strip().replace(',','.'))
#             personafisico.save()
#
#             persona.libretamilitar = row[14].strip()
#
#             persona.identificacioninstitucion = row[15].strip()
#             # tipo sangre
#             try:
#                 if row[16].strip() == '1':
#                     sangre = 5
#                 elif row[16].strip() == '2':
#                     sangre = 6
#                 elif row[16].strip() == '3':
#                     sangre = 3
#                 elif row[16].strip() == '4':
#                     sangre = 4
#                 elif row[16].strip() == '5':
#                     sangre = 7
#                 elif row[16].strip() == '6':
#                     sangre = 8
#                 elif row[16].strip() == '7':
#                     sangre = 1
#                 elif row[16].strip() == '8':
#                     sangre = 2
#                 else:
#                     sangre = 7
#             except:
#                 civil = 1
#             persona.sangre_id = sangre
#             # residencia
#             if row[17].strip() != '' and row[17].strip() != '0':
#                 persona.pais_id = int(row[17].strip())
#             if row[18].strip() != '' and row[18].strip() != '0':
#                 persona.provincia_id = int(row[18].strip())
#             if row[19].strip() != '' and row[19].strip() != '0':
#                 persona.canton_id = int(row[19].strip())
#             if row[20].strip() != '' and row[20].strip() != '0':
#                 persona.parroquia_id = int(row[20].strip())
#             # nacimiento
#             if row[21].strip() != '' and row[21].strip() != '0':
#                 persona.paisnacimiento_id = int(row[21].strip())
#             if row[22].strip() != '' and row[22].strip() != '0':
#                 persona.provincianacimiento_id = int(row[22].strip())
#             if row[23].strip() != '' and row[23].strip() != '0':
#                 persona.cantonnacimiento_id = int(row[23].strip())
#             if row[24].strip() != '' and row[24].strip() != '0':
#                 persona.parroquianacimiento_id = int(row[24].strip())
#
#             persona.anioresidencia = int(row[25].strip())
#
#             if Pais.objects.filter(pk=int(row[21].strip())).exists():
#                 pais1 = Pais.objects.filter(pk=int(row[21].strip()))[0]
#                 persona.nacionalidad = pais1.nacionalidad
#
#             # Servidor Carrera
#             try:
#                 if row[26].strip() == '<N>' or row[26].strip() == '':
#                     servidor = False
#                 else:
#                     servidor = True
#             except:
#                 servidor = False
#             persona.servidorcarrera = servidor
#
#             persona.regitrocertificacion = row[27].strip()
#
#             perfil = persona.mi_perfil()
#             discapacidad1 = False
#             if row[28].strip() == '<S>':
#                 discapacidad1 = True
#             perfil.tienediscapacidad = discapacidad1
#             perfil.carnetdiscapacidad = row[29].strip()
#             # tipo discapacidad
#             if int(row[30].strip()) != 0:
#                 perfil.tipodiscapacidad_id = int(row[30].strip())
#                 perfil.porcientodiscapacidad = float(row[31].strip().replace(',','.'))
#
#             if int(row[32].strip()) != 0:
#                 perfil.raza_id = int(row[32].strip())
#                 if int(row[33].strip()) != 0:
#                     perfil.nacionalidadindigena_id = int(row[33].strip())
#             perfil.save()
#             persona.sector = row[2].strip()
#             persona.direccion2 = row[35].strip()
#             persona.referencia = row[36].strip()
#             persona.num_direccion = row[37].strip()
#
#             persona.save()
#     n += 1

# ingreso de informacion academica
# workbook = xlrd.open_workbook(csv_filepathname)
# sheet = workbook.sheet_by_index(0)
# n = 1
# titulacion = Titulacion.objects.all().delete()
# for rowx in range(sheet.nrows):
#     print n
#     if n > 1:
#         row = sheet.row_values(rowx)
#         if n == 1271:
#             pass
#         if Persona.objects.filter(cedula=row[0].strip()).exists():
#             persona = Persona.objects.filter(cedula=row[0].strip())[0]
#             titulo = Titulo.objects.filter(pk=row[1].strip())[0]
#             areatitulo = None
#             if int(row[12].strip()) != 0:
#                 areatitulo = AreaTitulo.objects.filter(pk=row[12].strip())[0]
#             principal = False
#             if row[2].strip() == '<S>':
#                 principal = True
#
#             fechainicio = None
#             if row[4].strip() != '':
#                 fechainicio = convertirfecha2(row[4].strip())
#
#             fechaobtencion = None
#             if row[5].strip() != '':
#                 fechaobtencion = convertirfecha2(row[5].strip())
#
#             fechaegresado = None
#             if row[6].strip() != '':
#                 fechaegresado = convertirfecha2(row[6].strip())
#
#             areaconocimientotitulacion = None
#             subareaconocimientotitulacion = None
#             subareaespecificaconocimientotitulacion = None
#
#             if int(row[7].strip()) != 0:
#                 areaconocimientotitulacion = AreaConocimientoTitulacion.objects.filter(pk=int(row[7].strip()))[0]
#                 if int(row[8].strip()) != 0:
#                     subareaconocimientotitulacion = SubAreaConocimientoTitulacion.objects.filter(pk=int(row[8].strip()))[0]
#                     if int(row[9].strip()) != 0:
#                         subareaespecificaconocimientotitulacion = SubAreaEspecificaConocimientoTitulacion.objects.filter(pk=int(row[9].strip()))[0]
#
#             pais = None
#             canton = None
#             provincia = None
#             if row[10].strip() != '' and row[10].strip() != '0':
#                 pais = Pais.objects.filter(pk=int(row[10].strip()))[0]
#                 if row[11].strip() != '' and row[11].strip() != '0':
#                     canton = Canton.objects.filter(pk=int(row[11].strip()))[0]
#                     provincia = canton.provincia
#
#             educacionsuperior = False
#             if row[16].strip() != '' and row[16].strip() != '0':
#                 educacionsuperior = True
#                 institucioneducacionsuperior = InstitucionEducacionSuperior.objects.filter(pk=int(row[16].strip()))[0]
#                 colegio = None
#             else:
#                 institucioneducacionsuperior = None
#                 if row[15].strip() != '' and row[15].strip() != '0':
#                     colegio = Colegio.objects.filter(pk=int(row[15].strip()))[0]
#
#             anios = 0
#             if row[13].strip() != '':
#                 anios = int(row[13].strip())
#
#             semestre = 0
#             if row[14].strip() != '':
#                 semestre = int(row[14].strip())
#
#             estado = False
#             if row[17].strip() == 'TERMINADO':
#                 estado = True
#
#             titulacion = Titulacion(persona = persona,
#                                     titulo = titulo,
#                                     areatitulo = areatitulo,
#                                     principal = principal,
#                                     fechainicio = fechainicio,
#                                     fechaobtencion = fechaobtencion,
#                                     fechaegresado = fechaegresado,
#                                     registro = row[3].strip(),
#                                     areaconocimiento = areaconocimientotitulacion,
#                                     subareaconocimiento = subareaconocimientotitulacion,
#                                     subareaespecificaconocimiento = subareaespecificaconocimientotitulacion,
#                                     pais = pais,
#                                     provincia = provincia,
#                                     canton = canton,
#                                     # parroquia = models.ForeignKey(Parroquia, blank=True, null=True, verbose_name=u"Parroquia")
#                                     educacionsuperior = educacionsuperior,
#                                     institucion = institucioneducacionsuperior,
#                                     colegio = colegio,
#                                     anios = anios,
#                                     semestres = semestre,
#                                     cursando = estado,
#                                     verificado = True)
#             titulacion.save()
#     n += 1
# print 'listo'


# capacitacion personal
# workbook = xlrd.open_workbook(csv_filepathname)
# sheet = workbook.sheet_by_index(0)
# n = 1
# Capacitacion.objects.all().delete()
# for rowx in range(sheet.nrows):
#     print n
#     if n > 1:
#         row = sheet.row_values(rowx)
#         # if n == 1271:
#         #     pass
#         if Persona.objects.filter(cedula=row[0].strip()).exists():
#             persona = Persona.objects.filter(cedula=row[0].strip())[0]
#             auspiciante = ''
#             if row[5].strip() != '' and row[5].strip() != '0':
#                 auspiciante = row[5].strip()
#
#             pais = None
#             provincia = None
#             canton = None
#             parroquia = None
#             if row[6].strip() != '' and row[6].strip() != '0':
#                 pais = Pais.objects.filter(pk=int(row[6].strip()))[0]
#                 if row[7].strip() != '' and row[7].strip() != '0':
#                     canton = Canton.objects.filter(pk=int(row[7].strip()))[0]
#                     provincia = canton.provincia
#             fechainicio = None
#             if row[8].strip() != '':
#                 fechainicio = convertirfecha2(row[8].strip())
#             fechafin = None
#             if row[9].strip() != '':
#                 fechafin = convertirfecha2(row[9].strip())
#             horas = 0
#             if row[10].strip() != '':
#                 horas = float(row[10].strip())
#
#             tipocurso = None
#             if row[3].strip() != '' and row[3].strip() != '0':
#                 if TipoCurso.objects.filter(pk=int(row[3].strip())).exists():
#                     tipocurso = TipoCurso.objects.filter(pk=int(row[3].strip()))[0]
#
#             tipocertificacion = None
#             if row[4].strip() != '' and row[4].strip() != '0':
#                 if TipoCertificacion.objects.filter(pk=int(row[4].strip())).exists():
#                     tipocertificacion = TipoCertificacion.objects.filter(pk=int(row[4].strip()))[0]
#
#             capacitacion = Capacitacion(persona = persona,
#                                         institucion = row[1].strip(),
#                                         nombre = row[2].strip(),
#                                         tipocurso = tipocurso,
#                                         tipocertificacion = tipocertificacion,
#                                         auspiciante = auspiciante,
#                                         pais = pais,
#                                         provincia = provincia,
#                                         canton = canton,
#                                         parroquia = parroquia,
#                                         fechainicio = fechainicio,
#                                         fechafin = fechafin,
#                                         horas = horas,
#                                         expositor = row[11].strip(),
#                                         tiempo = row[12].strip(),
#                                         verificado = True)
#             capacitacion.save()
#
#     n += 1
# print 'listo'


# # experiencia
# workbook = xlrd.open_workbook(csv_filepathname)
# sheet = workbook.sheet_by_index(0)
# n = 1
# ExperienciaLaboral.objects.all().delete()
# for rowx in range(sheet.nrows):
#     print n
#     if n > 1:
#         row = sheet.row_values(rowx)
#         # if n == 1271:
#         #     pass
#         if Persona.objects.filter(cedula=row[0].strip()).exists():
#             persona = Persona.objects.filter(cedula=row[0].strip())[0]
#
#             pais = None
#             provincia = None
#             canton = None
#             parroquia = None
#             if row[5].strip() != '' and row[5].strip() != '0':
#                 canton = Canton.objects.filter(pk=int(row[5].strip()))[0]
#                 provincia = canton.provincia
#                 pais = provincia.pais
#
#             fechainicio = None
#             if row[6].strip() != '':
#                 fechainicio = convertirfecha2(row[6].strip())
#             fechafin = None
#             if row[7].strip() != '':
#                 fechafin = convertirfecha2(row[7].strip())
#
#             motivosalida = None
#             if row[8].strip() != '' and row[8].strip() != '0':
#                 motivosalida = MotivoSalida.objects.filter(pk=int(row[8].strip()))[0]
#
#             otroregimenlaboral = None
#             if row[9].strip() != '' and row[9].strip() != '0':
#                 otroregimenlaboral = OtroRegimenLaboral.objects.filter(pk=int(row[9].strip()))[0]
#
#             dedicacionlaboral = None
#             if row[11].strip() != '' and row[11].strip() != '0':
#                 dedicacionlaboral = DedicacionLaboral.objects.filter(pk=int(row[11].strip()))[0]
#
#             actividadlaboral = None
#             if row[12].strip() != '' and row[12].strip() != '0':
#                 actividadlaboral = ActividadLaboral.objects.filter(pk=int(row[12].strip()))[0]
#
#             experiencia = ExperienciaLaboral(persona = persona,
#                                             tipoinstitucion = int(row[1].strip()),
#                                             institucion = row[2].strip(),
#                                             cargo = row[3].strip(),
#                                             departamento = row[4].strip(),
#                                             pais = pais,
#                                             provincia = provincia,
#                                             canton = canton,
#                                             parroquia = parroquia,
#                                             fechainicio = fechainicio,
#                                             fechafin = fechafin,
#                                             motivosalida = motivosalida,
#                                             regimenlaboral = otroregimenlaboral,
#                                             horassemanales = int(row[10].strip()),
#                                             dedicacionlaboral = dedicacionlaboral,
#                                             actividadlaboral = actividadlaboral,
#                                             observaciones = row[13].strip(),
#                                             verificado = True)
#             experiencia.save(usuario=persona.usuario)
#     n += 1
# print 'listo'



# idiomas
# workbook = xlrd.open_workbook(csv_filepathname)
# sheet = workbook.sheet_by_index(0)
# n = 1
# IdiomaDomina.objects.all().delete()
# for rowx in range(sheet.nrows):
#     print n
#     if n > 1:
#         row = sheet.row_values(rowx)
#         # if n == 1271:
#         #     pass
#         if Persona.objects.filter(cedula=row[0].strip()).exists():
#             persona = Persona.objects.filter(cedula=row[0].strip())[0]
#             idioma = None
#             if row[1].strip() != '':
#                 idioma = Idioma.objects.filter(nombre=row[1].strip())[0]
#
#             idiomasdomina = IdiomaDomina(persona=persona,
#                                          idioma=idioma,
#                                          escritura=int(row[2].strip()),
#                                          lectura=int(row[3].strip()),
#                                          oral=int(row[4].strip()),
#                                          validado=True)
#             idiomasdomina.save()
#
#     n += 1
# print 'listo'

# # carga familiar
# workbook = xlrd.open_workbook(csv_filepathname)
# sheet = workbook.sheet_by_index(0)
# n = 1
# # PersonaDatosFamiliares.objects.all().delete()
# for rowx in range(sheet.nrows):
#     print n
#     if n > 1:
#         row = sheet.row_values(rowx)
#         # if not PersonaDatosFamiliares.objects.filter(identificacion=row[1].strip()).exists():
#         if Persona.objects.filter(cedula=row[0].strip()).exists():
#             persona = Persona.objects.filter(cedula=row[0].strip())[0]
#             fecha = None
#             if row[3].strip() != '':
#                 fecha = convertirfecha2(row[3].strip())
#
#             parentesco = None
#             if row[4].strip() != '':
#                 parentesco = ParentescoPersona.objects.filter(nombre=row[4].strip())[0]
#
#             discapacidad = False
#             if row[5].strip() == '<S>':
#                 discapacidad = True
#
#
#             cargafamiliar = PersonaDatosFamiliares(persona=persona,
#                                                     identificacion=row[1].strip(),
#                                                     nombre = row[2].strip(),
#                                                     fallecido = False,
#                                                     nacimiento = fecha,
#                                                     parentesco = parentesco,
#                                                     tienediscapacidad = discapacidad,
#                                                     telefono = '',
#                                                     telefono_conv = '',
#                                                     trabajo = '',
#                                                     convive = True,
#                                                     sustentohogar = False)
#             cargafamiliar.save()
#     n += 1
# print 'listo'



# # activo
# workbook = xlrd.open_workbook("activo.xlsx")
# sheet = workbook.sheet_by_index(0)
# n = 1
# ActivoFijo.objects.all().delete()
# for rowx in range(sheet.nrows):
#     print n
#     if n > 1:
#         row = sheet.row_values(rowx)
#         fechaingreso = None
#         if row[3].strip() != '':
#             fechaingreso = convertirfecha2(row[3].strip())
#         fechacomprobante = None
#         if row[12].strip() != '':
#             fechacomprobante = convertirfecha2(row[12].strip())
#         fechaultimadeprec = None
#         if row[13].strip() != '':
#             fechaultimadeprec = convertirfecha2(row[13].strip())
#
#         costo = 0
#         if row[5].strip() != '':
#             costo = float(row[5].strip().replace(',', '.'))
#
#         valorresidual = 0
#         if row[16].strip() != '':
#             valorresidual = float(row[16].strip().replace(',', '.'))
#
#         valorlibros = 0
#         if row[17].strip() != '':
#             valorlibros = float(row[17].strip().replace(',', '.'))
#
#         valordepreciacionacumulada = 0
#         if row[18].strip() != '':
#             valordepreciacionacumulada = float(row[18].strip().replace(',', '.'))
#
#         valordepreciacionanual = 0
#         if row[19].strip() != '':
#             valordepreciacionanual = float(row[19].strip().replace(',', '.'))
#
#         cuentacontable=None
#         if CuentaContable.objects.filter(cuenta=row[30].strip()).exists():
#             cuentacontable = CuentaContable.objects.filter(cuenta=row[30].strip())[0]
#
#         activo = ActivoFijo(id=int(row[0].strip()),
#                             codigogobierno=row[1].strip(),
#                             codigointerno=row[2].strip(),
#                             fechaingreso=fechaingreso,
#                             observacion=row[4].strip(),
#                             costo=costo,
#                             serie=row[6].strip(),
#                             descripcion=row[7].strip(),
#                             modelo=row[8].strip(),
#                             marca=row[9].strip(),
#                             tipocomprobante_id=int(row[10].strip()),
#                             numerocomprobante=row[11].strip(),
#                             fechacomprobante=fechacomprobante,
#                             fechaultimadeprec=fechaultimadeprec,
#                             deprecia=row[14].strip(),
#                             vidautil=int(row[15].strip()),
#                             valorresidual=valorresidual,
#                             valorlibros=valorlibros,
#                             valordepreciacionacumulada=valordepreciacionacumulada,
#                             valordepreciacionanual=valordepreciacionanual,
#                             subidogobierno=row[20].strip(),
#                             estructuraactivo=int(row[21].strip()),
#                             clasebien=int(row[22].strip()),
#                             catalogo_id=int(row[23].strip()),
#                             origeningreso_id=int(row[24].strip()),
#                             tipodocumentorespaldo_id=int(row[25].strip()),
#                             clasedocumentorespaldo_id=int(row[26].strip()),
#                             estado_id=int(row[27].strip()),
#                             statusactivo=True if row[32].strip() == 'True' else False,
#                             tipoproyecto_id=int(row[29].strip()),
#                             cuentacontable=cuentacontable,
#                             origenregistro=int(row[31].strip()))
#         activo.save(usuario=User.objects.get(pk=1))
#     n += 1
# print 'listo'




# activo bsjs
# workbook = xlrd.open_workbook("cbaja.xlsx")
# sheet = workbook.sheet_by_index(0)
# n = 1
# BajaActivo.objects.all().delete()
# for rowx in range(sheet.nrows):
#     print n
#     if n > 1:
#         row = sheet.row_values(rowx)
#         solicitante = Persona.objects.filter(cedula=row[3].strip())[0]
#         usuariobienentrega = Persona.objects.filter(cedula=row[8].strip())[0]
#         cbaja = BajaActivo(numero=row[0].strip(),
#                            fecha=convertirfecha2(row[1].strip()),
#                            tiposolicitud=int(row[2].strip()),
#                            solicitante=solicitante,
#                            oficio=row[4].strip(),
#                            fechaoficio=convertirfecha2(row[5].strip()),
#                            tipobaja_id=int(row[6].strip()),
#                            ubicacionbienentrega_id=int(row[7].strip()),
#                            usuariobienentrega=usuariobienentrega,
#                            usuariorecibe=row[9].strip(),
#                            cargorecibe=row[10].strip(),
#                            observacion=row[11].strip())
#         cbaja.save(usuario=User.objects.get(pk=1))
#
#     n += 1
# print 'listo'


# # detalle mantenimiento
# workbook = xlrd.open_workbook("detallemantenimiento.xlsx")
# DetalleMantenimiento.objects.all().delete()
# sheet = workbook.sheet_by_index(0)
# n = 1
# for rowx in range(sheet.nrows):
#     print n
#     if n > 1:
#         row = sheet.row_values(rowx)
#         detalletrasladomantenimiento = None
#         if row[0].strip() != '':
#             if DetalleTrasladoMantenimiento.objects.filter(pk=int(row[0].strip())).exists():
#                 detalletrasladomantenimiento = DetalleTrasladoMantenimiento.objects.filter(pk=int(row[0].strip()))[0]
#
#         aplicagarantia = False
#         if int(row[3].strip()) == 1:
#             aplicagarantia = True
#
#         costomanodeobra = 0
#         if row[4].strip() != '':
#             costomanodeobra = float(row[4].strip().replace(',', '.'))
#
#         costomanodereparacion = 0
#         if row[5].strip() != '':
#             costomanodereparacion = float(row[5].strip().replace(',', '.'))
#
#         detallemantenimiento = DetalleMantenimiento(detalletrasladomantenimiento = detalletrasladomantenimiento,
#                                                     mantenimientorealizar = row[1].strip(),
#                                                     mantenimientorealizado = row[2].strip(),
#                                                     aplicagarantia = aplicagarantia,
#                                                     costomanodeobra = costomanodeobra,
#                                                     costomanodereparacion = costomanodereparacion,
#                                                     facturamanodeobra = row[6].strip(),
#                                                     facturareparacion = row[7].strip(),
#                                                     fecharecepcion = convertirfecha2(row[8].strip()))
#         detallemantenimiento.save(usuario=User.objects.get(pk=1))
#     n +=1
# print 'listo'



#
# # CARACTERISTICA ACTIVO
# workbook = xlrd.open_workbook("caracteristicas.xlsx")
# sheet = workbook.sheet_by_index(0)
# n = 1
# for rowx in range(sheet.nrows):
#     print n
#     if n > 1:
#         row = sheet.row_values(rowx)
#         activo = ActivoFijo.objects.filter(pk=int(row[0].strip()))[0]
#         ######libro
#         #titulo
#         if int(row[1].strip()) == 1:
#             activo.titulo = row[2].strip()
#         #autor
#         if int(row[1].strip()) == 2:
#             activo.autor = row[2].strip()
#         #editorial
#         if int(row[1].strip()) == 3:
#             activo.editorial = row[2].strip()
#         #fecha edicion
#         if int(row[1].strip()) == 4:
#             activo.fechaedicion = convertirfecha2(row[2].strip())
#         #numero edicion
#         if int(row[1].strip()) == 5:
#             activo.numeroedicion = row[2].strip()
#         #clasificacion bibliografica
#         if int(row[1].strip()) == 6:
#             activo.clasificacionbibliografica = row[2].strip()
#
#         ######mueble
#         #color
#         if int(row[1].strip()) == 7:
#             activo.color_id = int(row[2].strip())
#         #material
#         if int(row[1].strip()) == 8:
#             activo.material = row[2].strip()
#         #dimensiones
#         if int(row[1].strip()) == 9:
#             activo.dimensiones = row[2].strip()
#
#         ######vehiculo
#         #clase vehiculo
#         if int(row[1].strip()) == 10:
#             activo.clasevehiculo_id = int(row[2].strip())
#         #tipo vehiculo
#         if int(row[1].strip()) == 11:
#             activo.tipovehiculo_id = int(row[2].strip())
#         #numero de motor
#         if int(row[1].strip()) == 12:
#             activo.numeromotor = row[2].strip()
#         #numero de chasis
#         if int(row[1].strip()) == 13:
#             activo.numerochasis = row[2].strip()
#         #año fabricacion
#         if int(row[1].strip()) == 14:
#             activo.aniofabricacion = int(row[2].strip())
#         #placa
#         if int(row[1].strip()) == 15:
#             activo.placa = row[2].strip()
#         #color1
#         if int(row[1].strip()) == 16:
#             activo.colorprimario_id = int(row[2].strip())
#         #color2
#         if int(row[1].strip()) == 17:
#             if row[2].strip() != '':
#                 activo.colorsecundario_id = int(row[2].strip())
#
#         activo.save(usuario=User.objects.get(pk=1))
#
#     n += 1
# print 'listo'


# # permiso institucional
# PermisoAprobacion.objects.all().delete()
# PermisoInstitucionalDetalle.objects.all().delete()
# PermisoInstitucional.objects.all().delete()
#
# workbook = xlrd.open_workbook("permiso2.xlsx")
# sheet = workbook.sheet_by_index(0)
# n = 1
# for rowx in range(sheet.nrows):
#     print n
#     if n > 1:
#         row = sheet.row_values(rowx)
#         persona = None
#         if row[1].strip() != '':
#             if Persona.objects.filter(cedula=row[1].strip()).exists():
#                 persona = Persona.objects.filter(cedula=row[1].strip())[0]
#
#         denominacionpuesto = None
#         if not DenominacionPuesto.objects.filter(descripcion=row[5].strip().upper()).exists():
#             denominacionpuesto = DenominacionPuesto(descripcion=row[5].strip(),
#                                                     codigo=str(n))
#             denominacionpuesto.save(usuario=User.objects.get(pk=1))
#         else:
#             denominacionpuesto = DenominacionPuesto.objects.filter(descripcion=row[5].strip().upper())[0]
#
#         departamento = None
#         if Departamento.objects.filter(nombre=row[6].strip().upper()).exists():
#             departamento = Departamento.objects.filter(nombre=row[6].strip().upper())[0]
#         else:
#             departamento = Departamento(nombre=row[6].strip().upper())
#             departamento.save(usuario=User.objects.get(pk=1))
#
#         tipopermiso = None
#         if row[7].strip() != '' and row[7].strip() != '0':
#             tipopermiso = TipoPermiso.objects.filter(pk=int(row[7].strip()))[0]
#
#         tipopermisodetalle = None
#         if row[8].strip() != '':
#             tipopermisodetalle = TipoPermisoDetalle.objects.filter(descripcion=row[8].strip())[0]
#
#         descuentovacaciones = False
#         if int(row[9].strip()) == 1:
#             descuentovacaciones = True
#
#         if persona:
#             if not PermisoInstitucional.objects.filter(secuencia=int(row[0].strip()),solicita=persona,fechasolicitud=convertirfecha2(row[2].strip())).exists():
#
#                 estadosolicitud = 1
#                 if int(row[10].strip()) == 2:
#                     estadosolicitud = 2
#                 elif int(row[10].strip()) == 4:
#                     estadosolicitud = 3
#                 elif int(row[10].strip()) == 3 or int(row[10].strip()) == 5:
#                     estadosolicitud = 4
#
#                 permiso = PermisoInstitucional(secuencia = int(row[0].strip()),
#                                                solicita = persona,
#                                                fechasolicitud = convertirfecha2(row[2].strip()),
#                                                tiposolicitud = int(row[3].strip()),
#                                                motivo = row[4].strip(),
#                                                denominacionpuesto = denominacionpuesto,
#                                                unidadorganica = departamento,
#                                                tipopermiso = tipopermiso,
#                                                tipopermisodetalle = tipopermisodetalle,
#                                                descuentovacaciones = descuentovacaciones,
#                                                estadosolicitud = estadosolicitud,
#                                                descripcionarchivo = row[11].strip())
#                 permiso.save(usuario=persona.usuario)
#
#                 if int(row[10].strip()) == 2:
#                     if Persona.objects.filter(cedula=row[12].strip()).exists():
#                         per1 = Persona.objects.get(cedula=row[12].strip())
#                     else:
#                         per1 = Persona.objects.get(pk=13416)
#                     permisoaprobacion = PermisoAprobacion(permisoinstitucional=permiso,
#                                                           observacion=row[4].strip(),
#                                                           aprueba=per1,
#                                                           fechaaprobacion=convertirfecha2(row[18].strip()),
#                                                           estadosolicitud=1)
#                     permisoaprobacion.save(usuario=per1.usuario)
#                 elif int(row[10].strip()) == 4:
#                     if Persona.objects.filter(cedula=row[12].strip()).exists():
#                         per1 = Persona.objects.get(cedula=row[12].strip())
#                     else:
#                         per1 = Persona.objects.get(pk=13416)
#
#                     permisoaprobacion = PermisoAprobacion(permisoinstitucional=permiso,
#                                                           observacion=row[4].strip(),
#                                                           aprueba=per1,
#                                                           fechaaprobacion=convertirfecha2(row[18].strip()),
#                                                           estadosolicitud=1)
#                     permisoaprobacion.save(usuario=per1.usuario)
#                     if Persona.objects.filter(cedula=row[13].strip()).exists():
#                         per2 = Persona.objects.get(cedula=row[13].strip())
#                     else:
#                         per2 = Persona.objects.get(pk=13416)
#                     permisoaprobacion = PermisoAprobacion(permisoinstitucional=permiso,
#                                                           observacion=row[4].strip(),
#                                                           aprueba=per2,
#                                                           fechaaprobacion=convertirfecha2(row[19].strip()),
#                                                           estadosolicitud=1)
#                     permisoaprobacion.save(usuario=per2.usuario)
#                 elif int(row[10].strip()) == 3:
#                     if Persona.objects.filter(cedula=row[12].strip()).exists():
#                         per1 = Persona.objects.get(cedula=row[12].strip())
#                     else:
#                         per1 = Persona.objects.get(pk=13416)
#                     permisoaprobacion = PermisoAprobacion(permisoinstitucional=permiso,
#                                                           observacion=row[4].strip(),
#                                                           aprueba=per1,
#                                                           fechaaprobacion=convertirfecha2(row[18].strip()),
#                                                           estadosolicitud=2)
#                     permisoaprobacion.save(usuario=per1.usuario)
#
#                 elif int(row[10].strip()) == 5:
#                     if Persona.objects.filter(cedula=row[12].strip()).exists():
#                         per1 = Persona.objects.get(cedula=row[12].strip())
#                     else:
#                         per1 = Persona.objects.get(pk=13416)
#                     permisoaprobacion = PermisoAprobacion(permisoinstitucional=permiso,
#                                                           observacion=row[4].strip(),
#                                                           aprueba=per1,
#                                                           fechaaprobacion=convertirfecha2(row[18].strip()),
#                                                           estadosolicitud=1)
#                     permisoaprobacion.save(usuario=per1.usuario)
#                     if Persona.objects.filter(cedula=row[13].strip()).exists():
#                         per2 = Persona.objects.get(cedula=row[13].strip())
#                     else:
#                         per2 = Persona.objects.get(pk=13416)
#                     permisoaprobacion = PermisoAprobacion(permisoinstitucional=permiso,
#                                                           observacion=row[4].strip(),
#                                                           aprueba=per2,
#                                                           fechaaprobacion=convertirfecha2(row[19].strip()),
#                                                           estadosolicitud=2)
#                     permisoaprobacion.save(usuario=per2.usuario)
#             else:
#                 permiso = PermisoInstitucional.objects.filter(secuencia=int(row[0].strip()), solicita=persona,fechasolicitud=convertirfecha2(row[2].strip()))[0]
#
#             horainicio = row[16].strip()
#             if row[16].strip() == '':
#                 horainicio = None
#
#             horafin = row[17].strip()
#             if row[17].strip() == '':
#                 horafin = None
#
#             permisoinstitucionaldetalle = PermisoInstitucionalDetalle(permisoinstitucional = permiso,
#                                                                       fechainicio = convertirfecha2(row[14].strip()),
#                                                                       fechafin = convertirfecha2(row[15].strip()),
#                                                                       horainicio = horainicio,
#                                                                       horafin = horafin)
#             permisoinstitucionaldetalle.save(usuario=persona.usuario)
#
#        # if
#     n += 1
# print 'listo'

# # persona
# workbook = xlrd.open_workbook("trabajador.xlsx")
# sheet = workbook.sheet_by_index(0)
# n = 1
# for rowx in range(sheet.nrows):
#     if n > 1:
#         row = sheet.row_values(rowx)
#         print row[0].strip()
#         if not Persona.objects.filter(cedula=row[0].strip()).exists():
#             apellidocompleto = row[2].strip()
#
#             apellido1 = apellidocompleto.split(' ')[0]
#             apellido2=''
#             try:
#                 apellido2 = apellidocompleto.split(' ')[1]
#             except:
#                 pass
#
#             persona = Persona(cedula=row[0].strip(),
#                               nombres=row[1].strip(),
#                               apellido1=apellido1,
#                               apellido2=apellido2,
#                               nacimiento=datetime.now().date(),
#                               sexo_id=SEXO_MASCULINO)
#
#             persona.save()
#             administrativo = Administrativo(persona=persona,
#                                             contrato='',
#                                             fechaingreso=datetime.now().date(),
#                                             activo=True)
#             administrativo.save()
#             username = calculate_username(persona)
#             generar_usuario(persona, username, ADMINISTRATIVOS_GROUP_ID)
#             if EMAIL_INSTITUCIONAL_AUTOMATICO:
#                 persona.emailinst = username + '@' + EMAIL_DOMAIN
#             persona.save()
#             persona.crear_perfil(administrativo=administrativo)
#             persona.mi_ficha()
#             persona.mi_perfil()
#             persona.datos_extension()
#
#     n += 1
# print 'listo'



# # # HOJA RUTA
# HojaRuta.objects.all().delete()
# workbook = xlrd.open_workbook("hojaruta.xlsx")
# sheet = workbook.sheet_by_index(0)
# n = 1
# for rowx in range(sheet.nrows):
#     if n > 1:
#         row = sheet.row_values(rowx)
#         print '%s  %s' % (n, row[8].strip())
#
#         solicitante = None
#         if row[8].strip() != '':
#             solicitante = Persona.objects.filter(cedula=row[8].strip())[0]
#
#         trabajador = None
#         if row[9].strip() != '':
#             trabajador = Persona.objects.filter(cedula=row[9].strip())[0]
#
#         departamento = None
#         if Departamento.objects.filter(nombre=row[2].strip().upper()).exists():
#             departamento = Departamento.objects.filter(nombre=row[2].strip().upper())[0]
#         else:
#             departamento = Departamento(nombre=row[2].strip().upper())
#             departamento.save(usuario=User.objects.get(pk=1))
#
#         hojaruta = HojaRuta(fecha = convertirfecha2(row[0].strip()),
#                             ubicacion = int(row[1].strip()),
#                             destinointerno = departamento,
#                             destinoexterno = row[3].strip(),
#                             actividad = row[4].strip(),
#                             horasalida = row[5].strip(),
#                             horaingreso = row[6].strip(),
#                             observacion = row[7].strip(),
#                             solicitante = solicitante,
#                             trabajador = trabajador)
#         hojaruta.save(usuario=trabajador.usuario)
#
#     n += 1
# print 'listo'

# # traspaso traspaso
# workbook = xlrd.open_workbook("traspaso.xlsx")
# # DetalleTraspasoActivo.objects.all().delete()
# # TraspasoActivo.objects.all().delete()
# usuariobien = ''
# custodiobien = ''
# area = 0
# numero = 0
# traspasoactivo = None
# sheet = workbook.sheet_by_index(0)
# n = 1
# for rowx in range(sheet.nrows):
#
#     cols = sheet.row_values(rowx)
#     traspasoactivo = TraspasoActivo.objects.filter(tipo=1,
#                                       fecha=convertirfecha2(cols[0]),
#                                       ubicacionbienrecibe_id=int(cols[3]),
#                                       custodiobienrecibe=Persona.objects.filter(cedula=cols[1].strip())[0],
#                                       usuariobienrecibe=Persona.objects.filter(cedula=cols[2].strip())[0]).update(fecha_creacion=convertirfechahora(cols[9].strip()))
#
#     # if ActivoFijo.objects.filter(id=int(cols[6])).exists():
#     #     traspasoactivo = TraspasoActivo.objects.filter(tipo=1,fecha=convertirfecha2(cols[0]),ubicacionbienrecibe_id=int(cols[3]),custodiobienrecibe=Persona.objects.filter(cedula=cols[1].strip())[0],usuariobienrecibe=Persona.objects.filter(cedula=cols[2].strip())[0])[0]
#     #     detalletraspasoactivo = DetalleTraspasoActivo.objects.filter(codigotraspaso = traspasoactivo,activo_id=int(cols[6])).update(fecha_creacion=convertirfechahora(cols[9].strip()))
#     print n
#     n += 1
#
# numero = 0
# n = 1
# traspasoactivo = None
# sheet = workbook.sheet_by_index(1)
# for rowx in range(sheet.nrows):
#     cols = sheet.row_values(rowx)
#     usuariocrea = Persona.objects.filter(cedula=cols[15].strip())[0]
#     if int(cols[0]) != numero:
#         numero = int(cols[0])
#         traspasoactivo = TraspasoActivo.objects.filter(tipo=2,fecha=convertirfecha2(cols[2]),fechaoficio=convertirfecha2(cols[6]),
#                                         solicitante=Persona.objects.filter(cedula=cols[4].strip())[0],oficio=cols[5],
#                                         ubicacionbienentrega_id=int(cols[7]),ubicacionbienrecibe_id=int(cols[8]),
#                                         custodiobienentrega=Persona.objects.filter(Q(cedula=cols[9].strip()) | Q(pasaporte=cols[9].strip()))[0],
#                                         custodiobienrecibe=Persona.objects.filter(Q(cedula=cols[10].strip()) | Q(pasaporte=cols[10].strip()))[0],
#                                         usuariobienentrega=Persona.objects.filter(Q(cedula=cols[11].strip()) | Q(pasaporte=cols[11].strip()))[0],
#                                         usuariobienrecibe=Persona.objects.filter(Q(cedula=cols[12].strip()) | Q(pasaporte=cols[12].strip()))[0],
#                                         observacion=cols[13],
#                                         responsablebienes_id=RESPONSABLE_BIENES_ID).update(fecha_creacion=convertirfechahora(cols[16].strip()))
#         # traspasoactivo.save(usuario=User.objects.get(pk=usuariocrea.usuario_id))
#     # if ActivoFijo.objects.filter(id=int(cols[14])).exists():
#     #     traspasoactivo = TraspasoActivo.objects.filter(tipo=2, fecha=convertirfecha2(cols[2]),
#     #                                                    fechaoficio=convertirfecha2(cols[6]),
#     #                                                    solicitante=Persona.objects.filter(cedula=cols[4].strip())[0],
#     #                                                    oficio=cols[5],
#     #                                                    ubicacionbienentrega_id=int(cols[7]),
#     #                                                    ubicacionbienrecibe_id=int(cols[8]),
#     #                                                    custodiobienentrega=Persona.objects.filter(Q(cedula=cols[9].strip()) | Q(pasaporte=cols[9].strip()))[0],
#     #                                                    custodiobienrecibe=Persona.objects.filter(Q(cedula=cols[10].strip()) | Q(pasaporte=cols[10].strip()))[0],
#     #                                                    usuariobienentrega=Persona.objects.filter(Q(cedula=cols[11].strip()) | Q(pasaporte=cols[11].strip()))[0],
#     #                                                    usuariobienrecibe=Persona.objects.filter(Q(cedula=cols[12].strip()) | Q(pasaporte=cols[12].strip()))[0],
#     #                                                    observacion=cols[13],responsablebienes_id=RESPONSABLE_BIENES_ID)[0]
#     #     detalletraspasoactivo = DetalleTraspasoActivo.objects.filter(codigotraspaso=traspasoactivo,
#     #                                                   activo_id=int(cols[14]),
#     #                                                   historico=False).update(fecha_creacion=convertirfechahora(cols[16].strip()))
#         # detalletraspasoactivo.activo.actualiza_responsable()
#         # DetalleTraspasoActivo.objects.filter(activo=detalletraspasoactivo.activo, codigotraspaso__fecha__lt=detalletraspasoactivo.codigotraspaso.fecha).update(historico=True)
#     print n
#     n += 1
# print 'listo'




# # traspaso traspaso
# workbook = xlrd.open_workbook("traspaso1.xlsx")
# # DetalleTraspasoActivo.objects.all().delete()
# # TraspasoActivo.objects.all().delete()
# usuariobien = ''
# custodiobien = ''
# area = 0
# numero = 0
# traspasoactivo = None
# sheet = workbook.sheet_by_index(0)
# n = 1
# for rowx in range(sheet.nrows):
#     cols = sheet.row_values(rowx)
#     usuariocrea = Persona.objects.filter(cedula=cols[8].strip())[0]
#     if cols[2].strip() != usuariobien or cols[1].strip() != custodiobien or int(cols[3]) != area:
#         numero += 1
#         traspasoactivo = TraspasoActivo(
#             tipo=1,
#             numero=numero,
#             fecha=convertirfecha2(cols[0]),
#             ubicacionbienrecibe_id=int(cols[3]),
#             custodiobienrecibe=Persona.objects.filter(cedula=cols[1].strip())[0],
#             usuariobienrecibe=Persona.objects.filter(cedula=cols[2].strip())[0],
#             observacion='',
#             fecha_creacion=convertirfechahora(cols[9].strip()),
#             responsablebienes_id=RESPONSABLE_BIENES_ID)
#
#         traspasoactivo.save(usuario=User.objects.get(pk=usuariocrea.usuario_id))
#     usuariobien = cols[2].strip()
#     custodiobien = cols[1].strip()
#     area = int(cols[3])
#     if ActivoFijo.objects.filter(id=int(cols[6])).exists():
#         detalletraspasoactivo = DetalleTraspasoActivo(
#             codigotraspaso = traspasoactivo,
#             activo_id=int(cols[6]),
#             historico = False,
#             fecha_creacion=convertirfechahora(cols[9].strip())
#         )
#         detalletraspasoactivo.save(usuario=User.objects.get(pk=usuariocrea.usuario_id))
#         detalletraspasoactivo.activo.actualiza_responsable()
#     print n
#     n += 1
# print 'listo'
#
# numero = 0
# n = 1
# traspasoactivo = None
# sheet = workbook.sheet_by_index(1)
# for rowx in range(sheet.nrows):
#     cols = sheet.row_values(rowx)
#     usuariocrea = Persona.objects.filter(cedula=cols[15].strip())[0]
#     if int(cols[0]) != numero:
#         numero = int(cols[0])
#         traspasoactivo = TraspasoActivo(tipo=2,
#                                         numero=numero,
#                                         fecha=convertirfecha2(cols[2]),
#                                         fechaoficio=convertirfecha2(cols[6]),
#                                         solicitante=Persona.objects.filter(cedula=cols[4].strip())[0],oficio=cols[5],
#                                         ubicacionbienentrega_id=int(cols[7]),ubicacionbienrecibe_id=int(cols[8]),
#                                         custodiobienentrega=Persona.objects.filter(Q(cedula=cols[9].strip()) | Q(pasaporte=cols[9].strip()))[0],
#                                         custodiobienrecibe=Persona.objects.filter(Q(cedula=cols[10].strip()) | Q(pasaporte=cols[10].strip()))[0],
#                                         usuariobienentrega=Persona.objects.filter(Q(cedula=cols[11].strip()) | Q(pasaporte=cols[11].strip()))[0],
#                                         usuariobienrecibe=Persona.objects.filter(Q(cedula=cols[12].strip()) | Q(pasaporte=cols[12].strip()))[0],
#                                         observacion=cols[13],
#                                         fecha_creacion=convertirfechahora(cols[16].strip()),
#                                         responsablebienes_id=RESPONSABLE_BIENES_ID)
#         traspasoactivo.save(usuario=User.objects.get(pk=usuariocrea.usuario_id))
#     if ActivoFijo.objects.filter(id=int(cols[14])).exists():
#         detalletraspasoactivo = DetalleTraspasoActivo(codigotraspaso=traspasoactivo,
#                                                       activo_id=int(cols[14]),
#                                                       historico=False,
#                                                       fecha_creacion=convertirfechahora(cols[16].strip()))
#         detalletraspasoactivo.save(usuario=User.objects.get(pk=usuariocrea.usuario_id))
#         # detalletraspasoactivo.activo.actualiza_responsable()
#         # DetalleTraspasoActivo.objects.filter(activo=detalletraspasoactivo.activo, codigotraspaso__fecha__lt=detalletraspasoactivo.codigotraspaso.fecha).update(historico=True)
#     print n
#     n += 1
# print 'listo'
#



# # numero custodio
# workbook = xlrd.open_workbook("numeroentrega.xlsx")
# sheet = workbook.sheet_by_index(0)
# n = 1
# for rowx in range(sheet.nrows):
#     print n
#     if n > 1:
#         row = sheet.row_values(rowx)
#         TraspasoActivo.objects.filter(Q(detalletraspasoactivo__activo__codigointerno=row[1].strip()) |
#                                       Q(detalletraspasoactivo__activo__codigogobierno=row[1].strip()),
#                                       tipo=1).update(numero=int(row[0].strip()))
#     n += 1

#
# workbook = xlrd.open_workbook("numero.xlsx")
# sheet = workbook.sheet_by_index(0)
# n = 1
# for rowx in range(sheet.nrows):
#     print n
#     if n > 1:
#         row = sheet.row_values(rowx)
#         traspaso = TraspasoActivo.objects.filter(tipo=1,numero=int(row[0].strip()))
#         acta = traspaso[0]
#         for t in traspaso[1:]:
#             for detalle in t.detalletraspasoactivo_set.all():
#                 detalle.codigotraspaso_id=acta.id
#                 detalle.save(usuario=acta.usuario_creacion)
#             t.delete()
#     n += 1


# # discapacitado
# workbook = xlrd.open_workbook("discapacitados.xlsx")
# sheet = workbook.sheet_by_index(0)
# n = 1
# for rowx in range(sheet.nrows):
#     print n
#     if n > 1:
#         row = sheet.row_values(rowx)
#         if Persona.objects.filter(cedula=row[0].strip()).exists():
#
#             persona = Persona.objects.filter(cedula=row[0].strip())[0]
#
#             if PerfilInscripcion.objects.filter(persona=persona).exists():
#                 perfilinscripcion = PerfilInscripcion.objects.filter(persona=persona)[0]
#                 perfilinscripcion.tienediscapacidad=True
#                 perfilinscripcion.tipodiscapacidad_id=int(row[1])
#                 perfilinscripcion.carnetdiscapacidad=row[2].strip()
#                 perfilinscripcion.porcientodiscapacidad=row[3]
#                 perfilinscripcion.save()
#             else:
#                 perfilinscripcion= persona.mi_perfil()
#                 perfilinscripcion.tienediscapacidad = True
#                 perfilinscripcion.tipodiscapacidad_id = int(row[1])
#                 perfilinscripcion.carnetdiscapacidad = row[2].strip()
#                 perfilinscripcion.porcientodiscapacidad = row[3]
#                 perfilinscripcion.save()
#         else:
#             print row[0].strip()
#     n += 1
# print 'listo'


# # discapacitado
# workbook = xlrd.open_workbook("actas_fecha.xlsx")
# sheet = workbook.sheet_by_index(0)
# n = 1
# for rowx in range(sheet.nrows):
#     print n
#     if n > 1:
#         row = sheet.row_values(rowx)
#         d = DetalleTraspasoActivo.objects.filter(codigotraspaso__tipo=1, activo_id=int(row[0].strip()))[0]
#         c=d.codigotraspaso
#         c.fecha=convertirfecha2(row[1].strip())
#         c.save(usuario=d.usuario_creacion)
#         d.activo.actualiza_responsable()
#     n += 1
# print 'listo'






# activo acta entrega
# workbook = xlrd.open_workbook("arreglocustodio.xlsx")
# sheet = workbook.sheet_by_index(0)
# n = 1
# for rowx in range(sheet.nrows):
#     print n
#     if n > 1:
#         row = sheet.row_values(rowx)
#         TraspasoActivo.objects.filter(tipo=1, fecha=convertirfechahora(row[0].strip()),ubicacionbienrecibe_id=int(row[1].strip()),usuariobienrecibe_id=int(row[2].strip()),custodiobienrecibe_id=int(row[3].strip())).update(numero=int(row[4].strip()))
#     n += 1

# # activo custodio
# workbook = xlrd.open_workbook("custodio.xlsx")
# sheet = workbook.sheet_by_index(0)
# n = 1
# lista = []
# for rowx in range(sheet.nrows):
#     if n > 1:
#         cols = sheet.row_values(rowx)
#         custodio = Persona.objects.filter(Q(cedula=cols[1].strip()) | Q(pasaporte=cols[1].strip()))[0]
#         responsable = Persona.objects.filter(Q(cedula=cols[2].strip()) | Q(pasaporte=cols[2].strip()))[0]
#
#         activo = ActivoFijo.objects.filter(pk=int(cols[0].strip()))[0]
#         if activo.responsable != responsable or activo.custodio != custodio:
#             lista.append([activo.id, responsable.id, custodio.id, activo.responsable.id, activo.custodio.id])
#     n += 1
#     print n
#
# for l in lista:
#     print '%s;%s;%s;%s;%s' % (str(l[0]),str(l[1]),str(l[2]),str(l[3]),str(l[4]))
#


# rubro = Rubro.objects.all()
# n = 1
#
# style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
# style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
# style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
# title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
# style1 = easyxf(num_format_str='D-MMM-YY')
# font_style = XFStyle()
# font_style.font.bold = True
# font_style2 = XFStyle()
# font_style2.font.bold = False
# wb = Workbook(encoding='utf-8')
# ws = wb.add_sheet('exp_xls_post_part')
# response = HttpResponse(mimetype='application/ms-excel')
# response['Content-Disposition'] = 'attachment; filename=exp_xls_post_part_' + random.randint(1, 10000).__str__() + '.xls'
#
# columns = [
#     (u"PERSONA", 6000),
#     (u"INSCRIPCION", 6000),
#     (u"IVA", 6000),
#     (u"FECHA EMISION", 6000),
#     (u"FECHA VENCIMIENTO", 6000),
#     (u"VALOR", 6000),
#     (u"RUBRO NOMBRE", 6000),
#     (u"TIPO", 6000)
#
# ]
# row_num = 1
# for col_num in xrange(len(columns)):
#     ws.write(row_num, col_num, columns[col_num][0], font_style)
#     ws.col(col_num).width = columns[col_num][1]
# row_num = 2
# for rubros in rubro:
#     print n
#     print '%s-%s-%s-%s-%s-%s-%s' % (rubros.inscripcion.persona_id,rubros.inscripcion_id,rubros.iva,rubros.fecha,rubros.fechavence,rubros.valor, rubros.nombre())
#
#
#     campo1 = rubros.inscripcion.persona_id
#     campo2 = rubros.inscripcion_id
#     campo3 = rubros.iva
#     campo4 = rubros.fecha
#     campo5 = rubros.fechavence
#     campo6 = rubros.valor
#     campo7 = rubros.nombre()
#     campo8 = rubros.tipo_exp()
#     ws.write(row_num, 0, campo1, font_style2)
#     ws.write(row_num, 1, campo2, font_style2)
#     ws.write(row_num, 2, campo3, font_style2)
#     ws.write(row_num, 3, campo4, style1)
#     ws.write(row_num, 4, campo5, style1)
#     ws.write(row_num, 5, campo6, font_style2)
#     ws.write(row_num, 6, campo7, font_style2)
#     ws.write(row_num, 7, campo8, font_style2)
#     row_num += 1
#     n +=1
# wb.save('rubros.xls')
# print 'listo'

# for discapacidad in Discapacidad.objects.all():
#     discapacidad.save()
#     print discapacidad.id
#
# for canton in Canton.objects.all():
#     if not canton.parroquia_set.exists():
#         parroquia = Parroquia(canton=canton,
#                               nombre=canton.nombre
#                               )
#         parroquia.save()
#         print "creado"



# # constatacion Fisica
# workbook = xlrd.open_workbook("cconstatacion.xlsx")
# sheet = workbook.sheet_by_index(0)
# n = 1
# ConstatacionFisica.objects.all().delete()
# for rowx in range(sheet.nrows):
#     print n
#     if n > 1:
#         if n == 40:
#             pass
#         row = sheet.row_values(rowx)
#         fechainicio = None
#         if row[1].strip() != '':
#             fechainicio = convertirfechahora(row[1].strip())
#             fechafin = None
#         if row[2].strip() != '':
#             fechafin = convertirfechahora(row[2].strip())
#
#         usuariobienes = None
#         if row[4].strip() != '':
#             usuariobienes = Persona.objects.filter(Q(cedula=row[4].strip()) | Q(pasaporte=row[4].strip()))[0]
#
#         constatacionfisica = ConstatacionFisica(id=int(row[0].strip()),
#                                                 numero=int(row[0].strip()),
#                                                 fechainicio = fechainicio,
#                                                 fechafin = fechafin,
#                                                 estado = int(row[3].strip()),
#                                                 usuariobienes = usuariobienes,
#                                                 ubicacionbienes_id = int(row[5].strip()),
#                                                 observacion = row[6].strip())
#         constatacionfisica.save(usuario=None)
#     n += 1







# # detalle constatacion
# workbook = xlrd.open_workbook("detalleconstatacion.xlsx")
# sheet = workbook.sheet_by_index(0)
# n = 1
# DetalleConstatacionFisica.objects.all().delete()
# for rowx in range(sheet.nrows):
#     print n
#     if n > 1:
#         row = sheet.row_values(rowx)
#         activo = None
#         if row[1].strip() != '':
#             if ActivoFijo.objects.filter(pk=int(row[1].strip())).exists():
#                 activo = ActivoFijo.objects.filter(pk=int(row[1].strip()))[0]
#
#         usuario = None
#         if row[5].strip() != '':
#             usuario = Persona.objects.filter(pk=int(row[5].strip()))[0]
#
#         ubicacionbienes = None
#         if row[6].strip() != '':
#             ubicacionbienes = Ubicacion.objects.filter(pk=int(row[6].strip()))[0]
#
#         requieretraspaso=False
#         if row[9].strip() == 'True':
#             requieretraspaso = True
#         if activo:
#             detalle = DetalleConstatacionFisica(codigoconstatacion_id=int(row[0].strip()),
#                                                 activo=activo,
#                                                 encontrado=row[2].strip(),
#                                                 enuso=row[3].strip(),
#                                                 perteneceusuario=row[4].strip(),
#                                                 usuariobienes=usuario,
#                                                 ubicacionbienes=ubicacionbienes,
#                                                 estadooriginal_id=int(row[7].strip()),
#                                                 estadoactual_id=int(row[8].strip()),
#                                                 requieretraspaso=requieretraspaso,
#                                                 observacion=row[10].strip())
#             detalle.save(usuario=User.objects.get(pk=1))
#     n += 1



# contratos
# workbook = xlrd.open_workbook("contratos.xlsx")
# sheet = workbook.sheet_by_index(0)
# n = 1
# ContratoPersonaDetalle.objects.all().delete()
# ContratoPersona.objects.all().delete()
# for rowx in range(sheet.nrows):
#     print n
#     if n > 1:
#         row = sheet.row_values(rowx)
#         persona = None
#         if row[0].strip() != '':
#             persona = Persona.objects.filter(cedula=row[0].strip())[0]
#
#             contratopersona = ContratoPersona(persona = persona,
#                                               contrato_id = int(row[2].strip()),
#                                               archivo = row[1].strip())
#             contratopersona.save(usuario=User.objects.get(pk=1))
#
#             c = 3
#             while c <= 56:
#                 if row[c].strip() != '' and row[c].strip() != '0':
#                     campos = int(row[c].strip())
#                     c += 1
#                     valor = row[c].strip()
#                     c += 1
#                     contratopersonadetalle = ContratoPersonaDetalle(contratopersona=contratopersona,
#                                                                     campos_id = campos,
#                                                                     valor = valor)
#                     contratopersonadetalle.save(usuario=User.objects.get(pk=1))
#                 #     para guardar departamento
#                     if campos == 2:
#                         if valor != '':
#                             if not Departamento.objects.filter(nombre=valor).exists():
#                                 departamento = Departamento(nombre=valor)
#                                 departamento.save(usuario=User.objects.get(pk=1))
#                 #     para guardar denominacionpuesto
#                     if campos == 3:
#                         if valor != '':
#                             if not DenominacionPuesto.objects.filter(descripcion=valor).exists():
#                                 denominacionpuesto = DenominacionPuesto(codigo = valor[:5].strip()+str(n),
#                                                                         descripcion=valor)
#                                 denominacionpuesto.save(usuario=User.objects.get(pk=1))
#
#                 else:
#                     c += 2
#     n += 1

#
# contratopersona = ContratoPersona.objects.all()
# for c in contratopersona:
#     c.archivo=c.archivo
#     c.save(usuario=User.objects.get(pk=1))
# print 'listo'





# # tipootrorubro
# workbook = xlrd.open_workbook("tipootrorubro.xlsx")
# sheet = workbook.sheet_by_index(0)
# n = 1
# TipoOtroRubro.objects.all().delete()
# for rowx in range(sheet.nrows):
#     print n
#     if n > 1:
#         row = sheet.row_values(rowx)
#         partida = None
#         if row[1].strip() != '':
#             if Partida.objects.filter(codigo=row[1].strip()).exists():
#                 partida = Partida.objects.filter(codigo=row[1].strip())[0]
#
#         tipootrorubro = TipoOtroRubro(nombre = row[0].strip(),
#                                       partida = partida)
#         tipootrorubro.save()
#     n += 1


# # constatacion usuario
# workbook = xlrd.open_workbook("constatacionuser.xlsx")
# sheet = workbook.sheet_by_index(0)
# n = 1
# for rowx in range(sheet.nrows):
#     print n
#     if n > 1:
#         row = sheet.row_values(rowx)
#         usuario_creacion = None
#         if row[1].strip() != '':
#             usuario_creacion = User.objects.filter(persona__cedula=row[1].strip())[0]
#
#         fecha_creacion = None
#         if row[2].strip() != '':
#             fecha_creacion = convertirfechahora(row[2].strip())
#
#         usuario_modificacion = None
#         if row[3].strip() != '':
#             usuario_modificacion = User.objects.filter(persona__cedula=row[3].strip())[0]
#
#         fecha_modificacion = None
#         if row[4].strip() != '':
#             fecha_modificacion = convertirfechahora(row[4].strip())
#
#         status = False
#         if int(row[5].strip()) == 1:
#             status = True
#
#         constatacionfisica = ConstatacionFisica.objects.filter(pk=int(row[0].strip()))[0]
#         constatacionfisica.usuario_creacion=usuario_creacion
#         constatacionfisica.fecha_creacion=fecha_creacion
#         constatacionfisica.usuario_modificacion=usuario_modificacion
#         constatacionfisica.fecha_modificacion=fecha_modificacion
#         constatacionfisica.status=status
#
#         constatacionfisica.save(usuario=usuario_creacion)
#     n +=1

# # detalle constatacion
# workbook = xlrd.open_workbook("detalleconstatacion2.xlsx")
# sheet = workbook.sheet_by_index(0)
# n = 1
# DetalleNoIdentificado.objects.all().delete()
# for rowx in range(sheet.nrows):
#     print n
#     if n > 1:
#         row = sheet.row_values(rowx)
#         detalle = DetalleNoIdentificado(codigoconstatacion_id=int(row[0].strip()),
#                                         codigobarra=row[1].strip(),
#                                         catalogobien_id=int(row[2].strip()),
#                                         serie=row[3].strip(),
#                                         descripcion=row[4].strip(),
#                                         modelo=row[5].strip(),
#                                         marca=row[6].strip(),
#                                         estado_id=int(row[7].strip()))
#         detalle.save(usuario=User.objects.get(pk=1))
#     n += 1



# traslado mantemiento
# workbook = xlrd.open_workbook("trasladomantenimiento.xlsx")
# sheet = workbook.sheet_by_index(0)
# n = 1
# TrasladoMantenimiento.objects.all().delete()
# DetalleTrasladoMantenimiento.objects.all().delete()
# for rowx in range(sheet.nrows):
#     print n
#     if n == 131:
#         pass
#     if n > 1:
#         row = sheet.row_values(rowx)
#         departamento = None
#         if row[6].strip() != '':
#             departamento = Departamento.objects.filter(nombre=row[6].strip())[0]
#
#         administradorcontrato = None
#         if row[11].strip() != '':
#             administradorcontrato = Persona.objects.filter(pk=int(row[11].strip()))[0]
#
#         asistentelogistica = None
#         if row[7].strip() != '':
#             asistentelogistica = Persona.objects.filter(pk=int(row[7].strip()))[0]
#
#         usuariobienes = None
#         if row[8].strip() != '':
#             usuariobienes = Persona.objects.filter(pk=int(row[8].strip()))[0]
#
#         taller = None
#         if row[10].strip() != '':
#             taller = TallerMantenimiento.objects.filter(pk=int(row[10].strip()))[0]
#
#         detalle = TrasladoMantenimiento(id=int(row[0].strip()),
#                                         numero=int(row[3].strip()),
#                                         fecha=convertirfecha2(row[5].strip()),
#                                         departamentosolicita=departamento,
#                                         asistentelogistica=asistentelogistica,
#                                         usuariobienes=usuariobienes,
#                                         observacion=row[9].strip(),
#                                         taller=taller,
#                                         administradorcontrato=administradorcontrato)
#         detalle.save(usuario=User.objects.get(pk=1))
#     n += 1



# # traslado mantemiento
# workbook = xlrd.open_workbook("detalletraslado.xlsx")
# sheet = workbook.sheet_by_index(0)
# n = 1
# for rowx in range(sheet.nrows):
#     print n
#     if n == 239:
#         pass
#     if n > 1:
#         row = sheet.row_values(rowx)
#
#         traslado = None
#         if row[0].strip() != '':
#             traslado = TrasladoMantenimiento.objects.filter(pk=int(row[0].strip()))[0]
#
#         activo = None
#         if row[1].strip() != '':
#             if ActivoFijo.objects.filter(pk=int(row[1].strip())).exists():
#                 activo = ActivoFijo.objects.filter(pk=int(row[1].strip()))[0]
#
#         if activo:
#             detalle = DetalleTrasladoMantenimiento(codigotraslado=traslado,
#                                                    activo=activo,
#                                                    observacion=row[2].strip())
#             detalle.save(usuario=User.objects.get(pk=1))
#     n += 1


# for activo in ActivoFijo.objects.all():
#     usuario = TraspasoActivo.objects.filter(detalletraspasoactivo__activo=activo, detalletraspasoactivo__historico=False).distinct()[0].usuariobienrecibe
#     custodio = TraspasoActivo.objects.filter(detalletraspasoactivo__activo=activo, detalletraspasoactivo__historico=False).distinct()[0].custodiobienrecibe
#     ubicacion = TraspasoActivo.objects.filter(detalletraspasoactivo__activo=activo, detalletraspasoactivo__historico=False).distinct()[0].ubicacionbienrecibe
#     activo.custodio = custodio
#     activo.ubicacion = ubicacion
#     activo.responsable = usuario
#     activo.save(usuario=User.objects.get(pk=1))
#     print activo



# def maximo(lista):
#     a = 0
#     elm = None
#     for e in lista:
#         if e[1] > a:
#             a = e[1]
#             elm = e
#     return elm
#
# problemas = 0
# for m in Matricula.objects.filter(nivel__periodo__id=32):
#     paralelos = []
#
#     if m.nivelmalla.id > 0:
#         for ma in m.materiaasignada_set.all():
#             p = ma.materia.paralelomateria
#             carrera = p.nombre.split('_')[1]
#             if carrera == 'EI':
#                 carreraid = [7]
#             elif carrera == 'EIB':
#                 carreraid = [11]
#             elif carrera == 'EB':
#                 carreraid = [8, 9, 10]
#             elif carrera == 'EIN':
#                 carreraid = [13]
#             else:
#                 carreraid = [0]
#             if m.inscripcion.carrera.id in carreraid:
#                 c = m.materiaasignada_set.filter(materia__paralelomateria=p).count()
#                 if [p, c] not in paralelos:
#                     paralelos.append([p, c])
#
#         # print m.id
#         # print paralelos
#         # print maximo(paralelos, m)
#         if not len(paralelos):
#             print "no tiene materia seleccionada en su carrera"
#         else:
#             print paralelos
#             print maximo(paralelos)
#         if len(paralelos) >= 1:
#             miparalelo = maximo(paralelos)[0]
#             for ot in m.materiaasignada_set.exclude(materia__paralelomateria=miparalelo):
#                 if Materia.objects.filter(paralelomateria=miparalelo, nivel__periodo__id=32, asignatura=ot.materia.asignatura).exists():
#                     mia = Materia.objects.filter(paralelomateria=miparalelo, nivel__periodo__id=32, asignatura=ot.materia.asignatura)[0]
#                     ot.materia = mia
#                     ot.save()
#                     print "cambio"
#                 else:
#                     print "no encontro"
#                     ot.delete()


# reseteo clave
# user = User.objects.filter(last_login=F('date_joined'))
# for u in user:
#     if Persona.objects.filter(usuario=u).exists():
#         persona = Persona.objects.filter(usuario=u)[0]
#         if persona.es_administrativo():
#             print persona
#             resetear_clave(persona)
#     else:
#         print 'no existe %s' % (u)
# print 'listo'

# # agente
# AgenteRiesgoRiesgo.objects.all().delete()
# AgenteRiesgo.objects.all().delete()
# workbook = xlrd.open_workbook("agente1.xlsx")
# sheet = workbook.sheet_by_index(0)
# n = 1
# for rowx in range(sheet.nrows):
#     print n
#     if n == 77:
#         pass
#     if n > 1:
#         row = sheet.row_values(rowx)
#         subgrupo = None
#         # subgrupo
#         if row[10] != '':
#             subgrupo = SubgrupoAgente.objects.filter(pk=int(row[10]))[0]
#
#         # riesgo
#         riesgo = 0
#         if row[11] != '':
#             riesgo = int(row[11])
#
#         if not AgenteRiesgo.objects.filter(codigo = row[4].strip()).exists():
#             agente = AgenteRiesgo(grupo_id = int(row[9]),
#                                   subgrupo = subgrupo,
#                                   apartado = row[13].strip(),
#                                   codigo = row[4].strip(),
#                                   descripcion = row[16].strip())
#             agente.save(usuario=User.objects.get(pk=1))
#         else:
#             agente = AgenteRiesgo.objects.filter(codigo = row[4].strip())[0]
#
#         if riesgo>0:
#             agenteriesgo = AgenteRiesgoRiesgo(agente=agente,
#                                              riesgo_id = riesgo,
#                                              medida = row[8].strip())
#             agenteriesgo.save(usuario=User.objects.get(pk=1))
#     n += 1
#



# for i in PerfilUsuario.objects.all():
#     if i.inscripcion.persona.perfilusuario_set.filter(inscripcion=i.inscripcion).count() > 1:
#         print i.inscripcion.persona
#
#
#
#     i.persona.crear_perfil(inscripcion=i)


# for rowx in range(sheet.nrows):
#     cols = sheet.row_values(rowx)
#     if not HistoricoRecordAcademico.objects.filter(inscripcion_id = int(cols[0]), asignatura_id = 554, fecha = convertirfecha(cols[2])).exists():
#         h = HistoricoRecordAcademico(inscripcion_id = int(cols[0]),
#                                      asignatura_id = 554,
#                                      nota = float(cols[1]),
#                                      asistencia = 100,
#                                      fecha = date(2015, 9, 25),
#                                      aprobada = True if cols[3] == 'TRUE' else False,
#                                      convalidacion = False,
#                                      creditos = 1,
#                                      horas = 40,
#                                      valida = True if cols[3] == 'TRUE' else False,
#                                      validapromedio = True,
#                                      observaciones = 'MIGRACION',
#                                      valoracioncalificacion_id=int(cols[6]))
#         h.save()
#         h.actualizar()
#         print h.id


#cierra todas las materias de los periodos
# for ma in MateriaAsignada.objects.all():
#     if ma.matricula.inscripcion.historicorecordacademico_set.filter(asignatura=ma.materia.asignatura, fecha=ma.materia.fin).exists():
#         print 'actualizo por materia'
#         h = ma.matricula.inscripcion.historicorecordacademico_set.filter(asignatura=ma.materia.asignatura, fecha=ma.materia.fin)[0]
#         h.fecha = ma.materia.fin
#         h.materiaregular = ma.materia
#         h.save()
#         h.actualizar()
#     elif ma.matricula.inscripcion.historicorecordacademico_set.filter(asignatura=ma.materia.asignatura, fecha=ma.matricula.nivel.periodo.fin).exists():
#         print 'actualizo por periodo'
#         h = ma.matricula.inscripcion.historicorecordacademico_set.filter(asignatura=ma.materia.asignatura, fecha=ma.matricula.nivel.periodo.fin)[0]
#         h.fecha = ma.materia.fin
#         h.materiaregular = ma.materia
#         h.save()
#         h.actualizar()
#     elif ma.matricula.inscripcion.historicorecordacademico_set.filter(asignatura=ma.materia.asignatura, fecha=ma.matricula.nivel.fin).exists():
#         print 'actualizo por nivel'
#         h = ma.matricula.inscripcion.historicorecordacademico_set.filter(asignatura=ma.materia.asignatura, fecha=ma.matricula.nivel.fin)[0]
#         h.fecha = ma.materia.fin
#         h.materiaregular = ma.materia
#         h.save()
#         h.actualizar()
#     else:
#         if ma.materia.cerrado:
#             print 'Creado'
#             ma.cierre_materia_asignada()
#         else:
#             print 'No creado'

#elimina repetidas
# for h in HistoricoRecordAcademico.objects.filter(materiaregular__isnull=False, id__lte=10107).order_by('-id'):
#     print h.id
#     for hi in h.recordacademico.historicorecordacademico_set.exclude(id=h.id):
#         if hi.nota == h.nota and hi.asistencia == h.asistencia:
#             hi.delete()
#             print "delete"
#     h.actualizar()

# actuliza creditos y horas
# for malla in Malla.objects.filter(carrera__id=22):
#     for am in malla.asignaturamalla_set.all():
#         RecordAcademico.objects.filter(inscripcion__inscripcionmalla__malla=malla, asignatura=am.asignatura).update(creditos=am.creditos, horas=am.horas, asignaturamalla=am)
#         HistoricoRecordAcademico.objects.filter(inscripcion__inscripcionmalla__malla=malla, asignatura=am.asignatura, aprobada=True).update(creditos=am.creditos, horas=am.horas, asignaturamalla=am)
#         print (am.id)

# for malla in Malla.objects.all():
#     for am in malla.modulomalla_set.all():
#         RecordAcademico.objects.filter(inscripcion__inscripcionmalla__malla=malla, asignatura=am.asignatura, aprobada=True).update(creditos=am.creditos, horas=am.horas, modulomalla=am)
#         HistoricoRecordAcademico.objects.filter(inscripcion__inscripcionmalla__malla=malla, asignatura=am.asignatura, aprobada=True).update(creditos=am.creditos, horas=am.horas, modulomalla=am)
#         print am.id





# CORREGIR NUMERO DE MATRICULAS EN EL RECORD
# from django.db.models import F
# for m in RecordAcademico.objects.annotate(cant=Count('historicorecordacademico')).filter(cant__gt=F('matriculas')).exclude(inscripcion__egresado__isnull=False):
#     print m.id
#     if not m.historicorecordacademico_set.filter(fecha=m.fecha).exists():
#         h = m.historicorecordacademico_set.all()[0]
#         h.actualizar()
#     else:
#         m.save()




# for m in Materia.objects.filter(nivel__periodo__id=11, modeloevaluativo__id=3):
#     for ma in m.asignados_a_esta_materia():
#         ma.actualiza_estado()
#         if m.cerrado:
#             ma.cierre_materia_asignada()
#         print ma.id


# for coordinadorcarrera in CoordinadorCarrera.objects.all():
#     print coordinadorcarrera.sede
#     print coordinadorcarrera.carrera
#     if Coordinacion.objects.filter(sede=coordinadorcarrera.sede, carrera__in=[coordinadorcarrera.carrera]).exists():
#         coordinacion = Coordinacion.objects.filter(sede=coordinadorcarrera.sede, carrera__in=[coordinadorcarrera.carrera])[0]
#         coordinadorcarrera.coordinacion = coordinacion
#         coordinadorcarrera.save()
#         print coordinadorcarrera.id
#     else:
#         coordinadorcarrera.delete()
#
#
# for pex in PersonaExtension.objects.all().order_by('-id'):
#     if pex.padre:
#         if not pex.datosfamiliares.filter(parentezco_id=1).exists():
#             parentesco = PersonaDatosFamiliares(
#                 parentezco_id = 1,
#                 nombre = pex.padre,
#                 fallecido = pex.fallecidopadre,
#                 cedula = pex.cedulapadre,
#                 edad = pex.edadpadre,
#                 estadocivil = pex.estadopadre,
#                 telefono = pex.telefpadre,
#                 educacion = pex.educacionpadre,
#                 profesion = pex.profesionpadre,
#                 trabajo = pex.trabajopadre
#             )
#             parentesco.save()
#             pex.datosfamiliares.add(parentesco)
#         if not pex.datosfamiliares.filter(parentezco_id=1).exists():
#             parentesco = PersonaDatosFamiliares(
#                 parentezco_id = 2,
#                 nombre = pex.madre,
#                 fallecido = pex.fallecidomadre,
#                 cedula = pex.cedulamadre,
#                 edad = pex.edadmadre,
#                 estadocivil = pex.estadomadre,
#                 telefono = pex.telefmadre,
#                 educacion = pex.educacionmadre,
#                 profesion = pex.profesionmadre,
#                 trabajo = pex.trabajomadre
#             )
#             parentesco.save()
#             pex.datosfamiliares.add(parentesco)
#         if not pex.datosfamiliares.filter(parentezco_id=1).exists():
#             parentesco = PersonaDatosFamiliares(
#                 parentezco_id = 13,
#                 nombre = pex.conyuge,
#                 fallecido = pex.fallecidoconyuge,
#                 cedula = pex.cedulaconyuge,
#                 edad = pex.edadconyuge,
#                 estadocivil = pex.estadoconyuge,
#                 telefono = pex.telefconyuge,
#                 educacion = pex.educacionconyuge,
#                 profesion = pex.profesionconyuge,
#                 trabajo = pex.trabajoconyuge
#             )
#             parentesco.save()
#             pex.datosfamiliares.add(parentesco)
#     print pex.id

#
# for c in Coordinacion.objects.all():
#     for ca in c.carrera.all():
#         for co in CoordinadorCarrera.objects.filter(coordinacion=c, carrera=ca).order_by('-id')[1:]:
#             co.delete()
#             print "Eliminado"






# Materia.objects.filter(id=3115, nivel__periodo__id=7).update(modeloevaluativo=ModeloEvaluativo.objects.get(pk=1))
# EvaluacionGenerica.objects.filter(materiaasignada__materia__nivel__periodo__id=7, materiaasignada__materia__id=3115, detallemodeloevaluativo__id=12).update(detallemodeloevaluativo=DetalleModeloEvaluativo.objects.get(pk=1))
# EvaluacionGenerica.objects.filter(materiaasignada__materia__nivel__periodo__id=7,materiaasignada__materia__id=3115, detallemodeloevaluativo__id=13).update(detallemodeloevaluativo=DetalleModeloEvaluativo.objects.get(pk=2))
# EvaluacionGenerica.objects.filter(materiaasignada__materia__nivel__periodo__id=7, materiaasignada__materia__id=3115,detallemodeloevaluativo__id=14).update(detallemodeloevaluativo=DetalleModeloEvaluativo.objects.get(pk=3))
# EvaluacionGenerica.objects.filter(materiaasignada__materia__nivel__periodo__id=7, materiaasignada__materia__id=3115,detallemodeloevaluativo__id=15).update(detallemodeloevaluativo=DetalleModeloEvaluativo.objects.get(pk=4))
# EvaluacionGenerica.objects.filter(materiaasignada__materia__nivel__periodo__id=7, materiaasignada__materia__id=3115,detallemodeloevaluativo__id=16).update(detallemodeloevaluativo=DetalleModeloEvaluativo.objects.get(pk=5))
# EvaluacionGenerica.objects.filter(materiaasignada__materia__nivel__periodo__id=7,materiaasignada__materia__id=3115, detallemodeloevaluativo__id=17).update(detallemodeloevaluativo=DetalleModeloEvaluativo.objects.get(pk=6))
# EvaluacionGenerica.objects.filter(materiaasignada__materia__nivel__periodo__id=7,materiaasignada__materia__id=3115, detallemodeloevaluativo__id=18).update(detallemodeloevaluativo=DetalleModeloEvaluativo.objects.get(pk=7))
# EvaluacionGenerica.objects.filter(materiaasignada__materia__nivel__periodo__id=7,materiaasignada__materia__id=3115, detallemodeloevaluativo__id=19).update(detallemodeloevaluativo=DetalleModeloEvaluativo.objects.get(pk=8))
# EvaluacionGenerica.objects.filter(materiaasignada__materia__nivel__periodo__id=7, materiaasignada__materia__id=3115, detallemodeloevaluativo__id=20).update(detallemodeloevaluativo=DetalleModeloEvaluativo.objects.get(pk=9))
# EvaluacionGenerica.objects.filter(materiaasignada__materia__nivel__periodo__id=7, materiaasignada__materia__id=3115, detallemodeloevaluativo__id=21).update(detallemodeloevaluativo=DetalleModeloEvaluativo.objects.get(pk=10))
# EvaluacionGenerica.objects.filter(materiaasignada__materia__nivel__periodo__id=7,materiaasignada__materia__id=3115, detallemodeloevaluativo__id=22).update(detallemodeloevaluativo=DetalleModeloEvaluativo.objects.get(pk=11))


# for i in Inscripcion.objects.all():
#     if Inscripcion.objects.filter(persona=i.persona, carrera=i.carrera).count() > 1:
#         print i


# for m in MensajeDestinatario.objects.all():
#     if MensajeDestinatario.objects.filter(mensaje=m.mensaje, destinatario=m.destinatario).count() > 1:
#         for e in MensajeDestinatario.objects.filter(mensaje=m.mensaje, destinatario=m.destinatario)[1:]:
#             e.delete()



# for r in RecordAcademico.objects.all():
#     if not r.historicorecordacademico_set.filter(fecha=r.fecha).exists():
#         r.actualizar()
#         print "actualizado"
#
# for ma in MateriaAsignada.objects.filter(materia__cerrado=True, materia__nivel__cerrado=True):
#     if HistoricoRecordAcademico.objects.filter(inscripcion=ma.matricula.inscripcion, asignatura=ma.materia.asignatura, fecha=ma.materia.fin).exists():
#         hist = HistoricoRecordAcademico.objects.filter(inscripcion=ma.matricula.inscripcion, asignatura=ma.materia.asignatura, fecha=ma.materia.fin)[0]
#         hist.materiaregular = ma.materia
#         hist.save()
#         hist.actualizar()
#     elif HistoricoRecordAcademico.objects.filter(inscripcion=ma.matricula.inscripcion, asignatura=ma.materia.asignatura, fecha=ma.materia.nivel.fin).exists():
#         hist = HistoricoRecordAcademico.objects.filter(inscripcion=ma.matricula.inscripcion, asignatura=ma.materia.asignatura, fecha=ma.materia.nivel.fin)[0]
#         hist.materiaregular = ma.materia
#         hist.fecha = ma.materia.fin
#         hist.save()
#         if RecordAcademico.objects.filter(inscripcion=ma.matricula.inscripcion, asignatura=ma.materia.asignatura, fecha=ma.materia.nivel.fin).exists():
#             hist = RecordAcademico.objects.filter(inscripcion=ma.matricula.inscripcion, asignatura=ma.materia.asignatura, fecha=ma.materia.nivel.fin)[0]
#             hist.materiaregular = ma.materia
#             hist.fecha = ma.materia.fin
#             hist.save()
#         print hist.inscripcion.id
#     else:
#         print hist.inscripcion.id







# for materia in Materia.objects.filter(modeloevaluativo__id=1, nivel__periodo__id=11):
#     materia.modeloevaluativo_id = 3
#     materia.save()
#     evaluaciones = EvaluacionGenerica.objects.filter(materiaasignada__materia=materia)
#     evaluaciones.delete()
#     for maa in materia.materiaasignada_set.all():
#         maa.evaluacion()
#         maa.notafinal = 0
#         maa.save()
#     if materia.cronogramaevaluacionmodelo_set.exists():
#         cronograma = materia.cronogramaevaluacionmodelo_set.all()[0]
#         cronograma.materias.remove(materia)
#     print "cambiado"



# nivel = Nivel.objects.get(pk=2)
# for i in Inscripcion.objects.filter(recordacademico__isnull=False):
#     if not Matricula.objects.filter(nivel=nivel, inscripcion=i).exists():
#         matricula = Matricula(nivel=nivel,
#                               inscripcion=i,
#                               fecha=date(2015, 5, 27),
#                               hora=datetime.now().time())
#         matricula.save()
#         malla = i.malla_inscripcion().malla
#         for rec in i.recordacademico_set.all():
#             materia = None
#             if malla.asignaturamalla_set.filter(asignatura=rec.asignatura).exists():
#                 am = malla.asignaturamalla_set.filter(asignatura=rec.asignatura)[0]
#                 if not nivel.materia_set.filter(identificacion=am.identificacion+'-'+str(am.id)).exists():
#                     materia = Materia(nivel=nivel,
#                                     asignatura=rec.asignatura,
#                                     asignaturamalla = am,
#                                     identificacion=am.identificacion+'-'+str(am.id),
#                                     alias = str(am.id),
#                                     horas = am.horas,
#                                     horassemanales = 0,
#                                     creditos = am.creditos,
#                                     inicio = nivel.inicio,
#                                     fin = nivel.fin,
#                                     fechafinasistencias = nivel.fin,
#                                     cerrado = True,
#                                     fechacierre = nivel.fin,
#                                     cupo = 100,
#                                     modeloevaluativo_id = 1,
#                                     validacreditos = True,
#                                     validapromedio = True)
#                     materia.save()
#                 else:
#                     materia = nivel.materia_set.filter(identificacion=am.identificacion+'-'+str(am.id))[0]
#             elif malla.modulomalla_set.filter(asignatura=rec.asignatura).exists():
#                 am = malla.modulomalla_set.filter(asignatura=rec.asignatura)[0]
#                 if not nivel.materia_set.filter(identificacion='MODULO'+str(am.id)).exists():
#                     materia = Materia(nivel=nivel,
#                                     asignatura=rec.asignatura,
#                                     modulomalla = am,
#                                     identificacion='MODULO'+str(am.id),
#                                     alias = str(am.id),
#                                     horas = 64,
#                                     horassemanales = 0,
#                                     creditos = 1.6,
#                                     inicio = nivel.inicio,
#                                     fin = nivel.fin,
#                                     fechafinasistencias = nivel.fin,
#                                     cerrado = True,
#                                     fechacierre = nivel.fin,
#                                     cupo = 100,
#                                     modeloevaluativo_id = 1,
#                                     validacreditos = True,
#                                     validapromedio = False)
#                     materia.save()
#                 else:
#                     materia = nivel.materia_set.filter(identificacion='MODULO'+str(am.id))[0]
#             if materia:
#                 materiaasignada = MateriaAsignada(matricula=matricula,
#                                                   materia=materia,
#                                                   notafinal=rec.nota,
#                                                   asistenciafinal=rec.asistencia,
#                                                   cerrado=True,
#                                                   fechacierre=nivel.fin,
#                                                   matriculas=1,
#                                                   observaciones='',
#                                                   estado_id= NOTA_ESTADO_APROBADO if rec.aprobada else NOTA_ESTADO_REPROBADO,
#                                                   fechaasignacion=matricula.fecha)
#                 materiaasignada.save()
#     print i.id



# for row in dataReader:
#     i = Inscripcion.objects.filter(persona__cedula=row[0])[0]
#     col = 3
#     for asi in [476, 480, 470, 472, 469, 554, 471, 529]:
#         if not i.historicorecordacademico_set.filter(asignatura__id=asi).exists():
#             h = HistoricoRecordAcademico(inscripcion = i,
#                                          asignatura_id = asi,
#                                          nota = float(row[col]),
#                                          asistencia = 100 if row[col+1] == 'A' else 0,
#                                          fecha = date(2015, 9, 25),
#                                          aprobada = True if row[col+1] == 'A' else False,
#                                          convalidacion = False,
#                                          creditos = 0,
#                                          horas = 0,
#                                          valida = True if row[col+1] == 'A' else False,
#                                          validapromedio = True if row[col+1] == 'A' else False,
#                                          observaciones = 'MIGRACION')
#             h.save()
#             h.actualizar()
#         col += 2
#         print row[0]


# for u in User.objects.filter(persona__perfilusuario__isnull=True):
#     links = [rel.get_accessor_name() for rel in u._meta.get_all_related_objects()]
#     for link in links:
#         objects = getattr(u, link).all()
#         if objects:
#             print objects
        # for object in objects:
        #     print object.name



# for l in LogEntry.objects.filter(user__persona__perfilusuario__isnull=True):
#     print l.id
#     if Persona.objects.filter(usuario=l.user).exists():
#         p = Persona.objects.filter(usuario=l.user)[0]
#         if Persona.objects.filter(cedula=p.cedula, perfilusuario__isnull=False).exists():
#             p = Persona.objects.filter(cedula=p.cedula, perfilusuario__isnull=False)[0]
#             l.user_id = p.usuario.id
#             l.save()
#             print "cambiado"



# for c in PersonaExtension.objects.filter(persona__perfilusuario__isnull=True):
#     if c.persona:
#         if Persona.objects.filter(cedula=c.persona.cedula, perfilusuario__isnull=False).exists():
#             p = Persona.objects.filter(cedula=c.persona.cedula, perfilusuario__isnull=False)[0]
#             if not p.personaextension_set.exists():
#                 c.persona = p
#                 c.save()
#                 print c.id


# for c in AgregacionEliminacionMaterias.objects.filter(responsable__perfilusuario__isnull=True):
#     if Persona.objects.filter(cedula=c.responsable.cedula, perfilusuario__isnull=False).exists():
#         p = Persona.objects.filter(cedula=c.responsable.cedula, perfilusuario__isnull=False)[0]
#         c.responsable = p
#         c.save()
#         print c.id

# for c in CVPersona.objects.filter(persona__perfilusuario__isnull=True):
#     if Persona.objects.filter(cedula=c.persona.cedula, perfilusuario__isnull=False).exists():
#         p = Persona.objects.filter(cedula=c.persona.cedula, perfilusuario__isnull=False)[0]
#         if not p.cvpersona_set.exists():
#             c.persona = p
#             c.save()
#             print c.id

# for c in ReservaDocumento.objects.filter(persona__perfilusuario__isnull=True):
#     if Persona.objects.filter(cedula=c.persona.cedula, perfilusuario__isnull=False).exists():
#         p = Persona.objects.filter(cedula=c.persona.cedula, perfilusuario__isnull=False)[0]
#         c.persona = p
#         c.save()
#         print c.id

# for c in PrestamoDocumento.objects.filter(persona__perfilusuario__isnull=True):
#     if Persona.objects.filter(cedula=c.persona.cedula, perfilusuario__isnull=False).exists():
#         p = Persona.objects.filter(cedula=c.persona.cedula, perfilusuario__isnull=False)[0]
#         c.persona = p
#         c.save()
#         print c.id

# for c in JustificacionAusenciaAsistenciaLeccion.objects.filter(persona__perfilusuario__isnull=True):
#     if Persona.objects.filter(cedula=c.persona.cedula, perfilusuario__isnull=False).exists():
#         p = Persona.objects.filter(cedula=c.persona.cedula, perfilusuario__isnull=False)[0]
#         c.persona = p
#         c.save()
#         print c.id

# for c in CoordinadorCarrera.objects.filter(persona__perfilusuario__isnull=True):
#     if Persona.objects.filter(cedula=c.persona.cedula, perfilusuario__isnull=False).exists():
#         p = Persona.objects.filter(cedula=c.persona.cedula, perfilusuario__isnull=False)[0]
#         c.persona = p
#         c.save()
#         print c.id



# for c in ConsultaBiblioteca.objects.filter(persona__perfilusuario__isnull=True):
#     if Persona.objects.filter(cedula=c.persona.cedula, perfilusuario__isnull=False).exists():
#         p = Persona.objects.filter(cedula=c.persona.cedula, perfilusuario__isnull=False)[0]
#         c.persona = p
#         c.save()
#         print c.id


# for r in DetalleInstrumentoEvaluacionIntegral.objects.filter(evaluador__perfilusuario__isnull=True):
#     print r.evaluador
#     print r.id
#     if r.evaluador:
#         if Persona.objects.filter(cedula=r.evaluador.cedula, perfilusuario__isnull=False).exists():
#             p = Persona.objects.filter(cedula=r.evaluador.cedula, perfilusuario__isnull=False)[0]
#             r.evaluador = p
#             r.save()
#             print "cambiado"



# for r in RespuestaEvaluacion.objects.filter(evaluador__perfilusuario__isnull=True):
#     print r.evaluador
#     print r.id
#     if r.evaluador:
#         if Persona.objects.filter(cedula=r.evaluador.cedula, perfilusuario__isnull=False).exists():
#             p = Persona.objects.filter(cedula=r.evaluador.cedula, perfilusuario__isnull=False)[0]
#             r.evaluador = p
#             r.save()
#             print "cambiado"


# for persona in Persona.objects.filter(inscripcion__isnull=True):
#     if persona.es_administrativo():
#         administrativo = persona.administrativo()
#         if not administrativo:
#             nuevo = Administrativo(persona=persona,
#                                    contrato='',
#                                    fechaingreso=datetime.now().date(),
#                                    activo=True)
#             nuevo.save()
#             print "Creando"
#         print persona.id




# for persona in Persona.objects.all().order_by('-id'):
#     crearperfil = True
#     print persona.id
#     du = persona.datos_unificados()
#     for pdu in du.persona.all():
#         if pdu.tiene_perfilusuario():
#             crearperfil = False
#             break
#     if crearperfil:
#         print "creando perfil"
#         for pdu in du.persona.all():
#             if pdu.es_estudiante():
#                 perfil = PerfilUsuario(persona=persona,
#                                        inscripcion=pdu.inscripcion())
#                 perfil.save()
#
#             if pdu.es_profesor():
#                 perfil = PerfilUsuario(persona=persona,
#                                        profesor=pdu.profesor())
#                 perfil.save()
#
#             if pdu.es_administrativo():
#                 perfil = PerfilUsuario(persona=persona,
#                                        administrativo=pdu.administrativo())
#                 perfil.save()


# for persona in Persona.objects.filter(perfilusuario__isnull=False).order_by('-id'):
#
#     for perfil in persona.mis_perfilesusuarios():
#         personaperfil = None
#         if perfil.es_administrativo():
#             administrativo = perfil.administrativo
#             personaperfil = administrativo.persona
#             if persona != personaperfil:
#                 administrativo.persona_id = persona.id
#                 administrativo.save()
#                 print "cambiado"
#         if perfil.es_profesor():
#             profesor = perfil.profesor
#             personaperfil = profesor.persona
#             if persona != personaperfil:
#                 profesor.persona_id = persona.id
#                 profesor.save()
#                 print "cambiado"
#         if perfil.es_estudiante():
#             estudiante = perfil.inscripcion
#             personaperfil = estudiante.persona
#             if persona != personaperfil:
#                 estudiante.persona_id = persona.id
#                 estudiante.save()
#                 print "cambiado"


# for persona in Persona.objects.filter(perfilusuario__isnull=False).order_by('-id'):
#     print persona.id
#     du = persona.datos_unificados()
#     for pdu in du.persona.all():
#         for gru in pdu.usuario.groups.all():
#             gru.user_set.add(persona.usuario)


# for i in Inscripcion.objects.all():
#     if i.persona.inscripcion_set.filter(carrera=i.carrera).count() > 1:
#         print i.persona.cedula



#actualizar nivel malla
# for nivel in Nivel.objects.all():
#     malla = nivel.malla
#     for materia in nivel.materia_set.all():
#         am = malla.asignaturamalla_set.filter(asignatura=materia.asignatura)
#         if am.exists():
#             materia.asignaturamalla = am[0]
#             materia.creditos = am[0].creditos
#             materia.horas = am[0].horas
#             materia.save()
#             print "AM - ok"
#         else:
#             print "Materia: no"
#     print "Nivel: " + str(nivel.id)


# for inscripcion in Inscripcion.objects.all().order_by('-id'):
#     inscripcion.actualizar_niveles_records()
#     inscripcion.actualizar_creditos()
#     inscripcion.actualizar_nivel()
#     print inscripcion.id


# for row in dataReader:
#     for da in Factura.objects.filter(numero=row[0], valida=False):
#         da.numero = da.numero + "C"
#         da.save()
#     print row[0]



# for row in dataReader:
#     for da in PersonaExtension.objects.filter(datosunificados__id=int(row[0]))[1:]:
#         da.delete()
#     print row[0]

# for row in dataReader:
#     for da in DocumentoCarrera.objects.filter(documento__id=int(row[0]))[1:]:
#         da.delete()
#     print row[0]

# for row in dataReader:
#     for da in PersonaConsultaMedica.objects.filter(documento__id=int(row[0]))[1:]:
#         da.delete()
#     print row[0]

# for row in dataReader:
#     for da in PersonaConsultaOdontologica.objects.filter(datosunificados__id=int(row[1]))[1:]:
#         da.delete()
#     print row[0]

# for row in dataReader:
#     for da in DatosUnificados.objects.filter(cedula=row[0])[1:]:
#         da.delete()
#     print row[0]

# for row in dataReader:
#     for da in Reporte.objects.filter(nombre=row[0])[1:]:
#         da.delete()
#     print row[0]

# for row in dataReader:
#     for da in AmbitoInstrumentoEvaluacion.objects.filter(instrumento__id=int(row[0]), ambito__id=int(row[1]))[1:]:
#         da.delete()
#     print row[0]


# for row in dataReader:
#     for da in DepositoInscripcion.objects.filter(inscripcion__id=int(row[0]), fecha=convertirfecha2(row[1]), cuentabanco__id=int(row[2]), referencia=row[3], deposito=(row[4] == 't'))[1:]:
#         da.referencia += "-"+str(da.id)
#
#         da.save()
#         print da.referencia
#     print row[0]

# for row in dataReader:
#     for da in EvaluacionProfesor.objects.filter(proceso__id=int(row[0]), instrumento__id=int(row[1]), profesor__id=int(row[2]), persona__id=int(row[3]), materia__id=int(row[4]))[1:]:
#         da.delete()
#     print row[0]

# for row in dataReader:
#     correcto = IndicadorEvaluacion.objects.filter(nombre=row[0])[0]
#     for da in IndicadorEvaluacion.objects.filter(nombre=row[0])[1:]:
#         da.indicadorambitoinstrumentoevaluacion_set.update(indicador=correcto)
#         da.delete()
#         print da.id

# for row in dataReader:
#     for perfil in InscripcionMalla.objects.filter(inscripcion__id=int(row[0]))[1:]:
#         perfil.delete()
#     print row[0]

# for row in dataReader:
#     for perfil in ProfesorDistributivoHoras.objects.filter(profesor__id=int(row[1]), periodo__id=int(row[0]))[1:]:
#         perfil.delete()
#     print row[0]

# for row in dataReader:
#     for eva in EvaluacionGenerica.objects.filter(materiaasignada__id=int(row[0]), detallemodeloevaluativo__id=int(row[1]))[1:]:
#         eva.delete()
#     print row[0]

# for row in dataReader:
#     for pm in ProfesorMateria.objects.filter(materia__id=int(row[0]), profesor__id=int(row[1]))[1:]:
#         pm.delete()
#     print row[0]

# for row in dataReader:
#     for perfil in PerfilInscripcion.objects.filter(datosunificados__id=int(row[0]))[1:]:
#         perfil.delete()
#     print row[0]

# for row in dataReader:
#     for du in ClienteFacturaInscripcion.objects.filter(inscripcion__id=int(row[0]))[1:]:
#         du.delete()
#     print row[0]

# for du in PagoCheque.objects.all():
#     if du.cuenta == '':
#         du.cuenta = "SIN REGISTRO - " + str(du.id)
#         du.save()


# for row in dataReader:
#     for du in MateriaAsignadaRetiro.objects.filter(materiaasignada=int(row[0]))[1:]:
#         du.delete()
#     print row[0]

# for row in dataReader:
#     horarios = Clase.objects.filter(materia_id=int(row[0]), turno_id=int(row[1]), dia=int(row[2]), fin=convertirfecha2(row[3]), inicio=convertirfecha2(row[4]))
#     primera = horarios[0]
#     primera.activo = True
#     primera.save()
#     for du in horarios[1:]:
#         du.leccion_set.all().update(clase=primera)
#         du.delete()
#     print row[0]

# for row in dataReader:
#     print "-->" + row[0]
#     fecha=convertirfecha2(row[0])
#     clase=Clase.objects.get(pk=int(row[1]))
#     for leccion in clase.leccion_set.filter(fecha=fecha)[1:]:
#         print leccion.id
#         leccion.delete()


# for row in dataReader:
#     for du in Incidencia.objects.filter(lecciongrupo__id=int(row[0]), contenido=row[1], tipo__id=int(row[2]))[1:]:
#         du.delete()
#     print row

# for row in dataReader:
#     for du in CierreSesionCaja.objects.filter(sesion__id=int(row[0]))[1:]:
#         du.delete()
#     print row

# for row in dataReader:
#     for du in ParticipanteActividadExtraCurricular.objects.filter(inscripcion__id=int(row[0]), actividad__id=int(row[1]))[1:]:
#         du.delete()
#     print du

# for row in dataReader:
#     for du in Pasantia.objects.filter(inscripcion__id=int(row[0]), inicio=convertirfecha2(row[1]), fin=convertirfecha2(row[2]))[1:]:
#         du.delete()
#     print row

# for row in dataReader:
#     rubro = Rubro.objects.get(pk=int(row[0]))
#     for du in DescuentoRecargoRubro.objects.filter(rubro=rubro)[1:]:
#         du.delete()
#     print rubro


# for row in dataReader:
#     for du in ClienteFactura.objects.filter(ruc=row[0])[1:]:
#         du.delete()
#     print row[0]

# for row in dataReader:
#     for du in PreguntasInscripcion.objects.filter(inscripcion__id=int(row[0]))[1:]:
#         du.delete()
#     print row[0]

# for row in dataReader:
#     for du in EvaluacionGenerica.objects.filter(materiaasignada__id=int(row[0]), detallemodeloevaluativo__id=int(row[1])).order_by('id')[1:]:
#         du.delete()
#     print row[0]

# for row in dataReader:
#     p =1
#     for du in Materia.objects.filter(nivel__id=int(row[0]), asignatura__id=int(row[1]), asignaturamalla__id=int(row[2]), identificacion=row[3], alias=row[4], paralelo=row[5]):
#         du.paralelo = 'A' + str(p)
#         du.save()
#         p += 1
#         print du
#     print row[0]

# for row in dataReader:
#     for du in FichaSocioEconomica.objects.filter(datosunificados__id=int(row[0]))[1:]:
#         du.delete()
#     print row[0]

# for row in dataReader:
#     for eva in EvaluacionProfesor.objects.filter(proceso_id=int(row[0]), instrumento_id=int(row[1]), profesor_id=int(row[2]), persona_id=int(row[3]), materia_id=int(row[4]))[1:]:
#         eva.delete()
#     print row


# for row in dataReader:
#     for pago in PagoTarjeta.objects.filter(tipo__id=int(row[0]), banco__id=int(row[1]), procesadorpago__id=int(row[2]), referencia=row[3], fecha=convertirfecha2(row[4]))[:1]:
#         pago.referencia = pago.referencia + '-' + str(pago.id)
#         pago.save()
#         print pago.id

# for row in dataReader:
#     for pago in PagoTransferenciaDeposito.objects.filter(deposito=(row[0] == 't'), cuentabanco__id=int(row[1]), fecha=convertirfecha2(row[2]), referencia=row[3], )[:1]:
#         pago.referencia = pago.referencia + '-' + str(pago.id)
#         pago.save()
#         print pago.id


# for row in :
#     pago.PagoTarjeta.objects.all()
#     pago.save()
#     print pago.id

# for row in dataReader:
#     especialidades = Especialidad.objects.filter(nombre=row[0])
#     for du in especialidades[1:]:
#         Inscripcion.objects.filter(especialidad=du).update(especialidad=especialidades[0])
#         du.delete()
#     print row[0]

# for row in dataReader:
#     for du in FacturaCancelada.objects.filter(factura_id=int(row[0]), sesion__isnull=True):
#         du.delete()
#     print row[0]

# for row in dataReader:
#     for du in DocumentosDeInscripcion.objects.filter(inscripcion__id=int(row[0]))[1:]:
#         du.delete()
#     print row[0]

# for row in dataReader:
#     for du in Documento.objects.filter(codigo=row[0])[1:]:
#         du.codigo += 'A'
#         du.save()
#     print row[0]


# for row in dataReader:
#     for du in Egresado.objects.filter(inscripcion__id=int(row[0]))[1:]:
#         du.delete()
#     print row[0]

# for row in dataReader:
#     for du in IdiomaDomina.objects.filter(datosunificados_id=int(row[0]), idioma_id=int(row[1]))[1:]:
#         du.delete()
#     print row[0]

# for row in dataReader:
#     for du in InscripcionNivel.objects.filter(inscripcion_id=int(row[0]))[1:]:
#         du.delete()
#     print row[0]

# for row in dataReader:
#     for du in InscripcionTesDrive.objects.filter(inscripcion_id=int(row[0]))[1:]:
#         du.delete()
#     print row[0]

# for row in dataReader:
#     for du in MatriculaCursoEscuelaComplementaria.objects.filter(inscripcion__id=int(row[0]), curso__id=int(row[1]))[1:]:
#         du.delete()
#     print row[0]

# for row in dataReader:
#     for du in NivelExtension.objects.filter(nivel__id=int(row[0]))[1:]:
#         du.delete()
#     print row[0]

# for row in dataReader:
#     for du in JustificacionAusenciaAsistenciaLeccion.objects.filter(asistencialeccion_id=int(row[0]))[1:]:
#         du.delete()
#     print row[0]

# for row in dataReader:
#     lg = LeccionGrupo.objects.get(pk=int(row[0]))
#     lg.turno_id = int(row[3])
#     lg.save()
#     print row[0]

# for row in dataReader:
#     for du in DatosUnificados.objects.filter(cedula=row[0])[1:]:
#         du.delete()
#     print row[0]


# arreglo duplicados record
# linea = 1
# for row in dataReader:
#     inscripcion = Inscripcion.objects.get(pk=int(row[0]))
#     if inscripcion.recordacademico_set.filter(asignatura__id=int(row[1])).exists():
#         print "arreglando"
#         for record in inscripcion.recordacademico_set.filter(asignatura__id=int(row[1])).order_by('-aprobada', 'nota', '-fecha')[1:]:
#             record.delete()
#             print "eliminado"
#     linea +=1
#     print linea


#arrglo lecciongrupo
# for row in dataReader:
#     lecciongrupo = LeccionGrupo.objects.get(pk=int(row[1]))
#     turno = Turno.objects.get(pk=row[3])
#     lecciongrupo.turno = turno
#     lecciongrupo.save()
#     print lecciongrupo.id


# for lg in LeccionGrupo.objects.all():
#     for le in lg.mis_leciones():
#         if le.fecha != lg.fecha:
#             print lg.id




#arreglo historico
# linea = 1
# for row in dataReader:
#     inscripcion = Inscripcion.objects.get(pk=int(row[0]))
#     print inscripcion
#     for record in inscripcion.historicorecordacademico_set.filter(asignatura__id=int(row[1]), fecha=convertirfecha2(row[2])).order_by('-aprobada', 'nota', '-fecha')[1:]:
#         record.delete()
#         print "eliminado"
#     linea +=1
#     print linea

#becarios
# linea = 1
# for row in dataReader:
#     if not Inscripcion.objects.filter(persona__cedula__icontains=row[1], carrera__id=int(row[12])).exists():
#         inscripcion = Inscripcion.objects.filter(persona__cedula__icontains=row[1], carrera__id=int(row[12]))[0]
#         if not inscripcion.inscripcionbecario_set.exists():
#             becario = InscripcionBecario(inscripcion=inscripcion,
#                                          tipobeca_id=int(row(2)),
#                                          porciento=0,
#                                          montomensual=0,
#                                          cantidadmeses=int(row[11]),
#                                          montobeneficio=float(row[10]),
#                                          motivo='BECAS MAYO/SEPTIEMBRE',
#                                          fecha=datetime.now().date(),
#                                          activo=True)
#             becario.save()
#         else:
#             becario = inscripcion.inscripcionbecario_set.all()[0]
#         if inscripcion.matricula_set.filter(nivel__periodo__id=5).exists():
#             matricula =  inscripcion.matricula_set.filter(nivel__periodo__id=5)[0]
#             if not matricula.becado:
#                 matricula.becado = True
#                 matricula.tipobeca = becario.tipobeca
#                 matricula.montobeneficio = becario.montobeneficio
#
#
#     linea += 1

# print "Comenzo"
# linea = 1
# grupo = 1
# records = RecordAcademico.objects.all().annotate(historicos=Count('historicorecordacademico')).filter(historicos__gt=1).order_by('-id')
# print records.count()
# for record in records:
#     for repetido in record.historicorecordacademico_set.order_by('fecha').values('fecha').annotate(dcount=Count('fecha')):
#         if repetido['dcount'] > 1:
#             for eliminar in record.historicorecordacademico_set.filter(fecha=repetido['fecha']).order_by('-nota')[1:]:
#                 eliminar.delete()
#             print record.asignatura
#             print record.inscripcion
#     linea += 1
#     if linea == 1000:
#         grupo += 1
#         linea = 1
#         print "NO - " + str(grupo)

# linea = 0
# for matricula in Matricula.objects.filter(nivel__id__in=[42, 43]).distinct():
#     materias = matricula.materiaasignada_set.exclude(materia__asignatura__modulo=True).count()
#     materiasnivel = matricula.materiaasignada_set.filter(materia__nivel=matricula.nivel).exclude(materia__asignatura__modulo=True).count()
#     if materias != materiasnivel:
#         print matricula.inscripcion.persona.cedula + " " +str(materias) + " " +str(materiasnivel) + " " + matricula.inscripcion.carrera.alias
#         linea += 1
# print linea
# linea = 0
# for row in dataReader:
#     if Inscripcion.objects.filter(id=int(row[4])).exists():
#         per = Inscripcion.objects.get(pk=int(row[4])).persona
#         usua = per.usuario.username
#         usun = calculate_username(per)
#         usuario = per.usuario
#         usuario.username = usun
#         usuario.save()
#         resetear_clave(per)
#         per.cambiar_clave()
#         print usua + " ----- " + usun



# linea = 0
# for row in dataReader:
#     if MateriaAsignada.objects.filter(id=int(row[0])).exists():
#         materia = Materia.objects.filter(materiaasignada__id=int(row[0])).distinct()[0]
#         if materia.cerrado:
#             materia.cerrado = False
#             materia.save()
#             materia.materiaasignada_set.update(cerrado=False)
#         if materia.modeloevaluativo_id == 1:
#             cronograma = CronogramaEvaluacionModelo.objects.get(pk=7)
#         else:
#             cronograma = CronogramaEvaluacionModelo.objects.get(pk=6)
#         cronograma.materias.add(materia)
#         ma = MateriaAsignada.objects.get(pk=int(row[0]))
#         if row[1] == 'x':
#             ma.sinasistencia = True
#             ma.save()
#     linea += 1
#     print linea

# for materiaasignada in MateriaAsignada.objects.filter(asistenciafinal__lt=70, sinasistencia=False).order_by('-id'):
#     evaluacion = materiaasignada.evaluacion()
#     modeloevaluativo = materiaasignada.materia.modeloevaluativo
#     exec modeloevaluativo.logicamodelo
#     calculo_modelo_evaluativo(materiaasignada)
#     materiaasignada.notafinal = round(materiaasignada.notafinal, modeloevaluativo.notafinaldecimales)
#     materiaasignada.save()
#     camposdependientes = []
#     encurso = True
#     for campomodelo in modeloevaluativo.campos():
#         if campomodelo.actualizaestado and materiaasignada.valor_nombre_campo(campomodelo.nombre) > 0:
#             encurso = False
#     if not encurso:
#         materiaasignada.actualiza_estado()
#     else:
#         materiaasignada.estado_id = NOTA_ESTADO_EN_CURSO
#         materiaasignada.save()
#     print materiaasignada.id


# for ma in Materia.objects.all().order_by('-id'):
#     ma.modeloevaluativo_id = 1
#     ma.save()
#     print ma.id
#
# for ma in MateriaAsignada.objects.all().order_by('-id'):
#     ma.evaluacion()
#     print ma.id
#
#
# for ma in MateriaAsignada.objects.all().order_by('-id'):
#     evalua = ma.evaluacion_libertad()
#     ne = ma.evaluacion()
#     campo1 = ma.campo('N1')
#     campo1.valor = evalua.n1
#     campo1.save()
#     campo2 = ma.campo('N2')
#     campo2.valor = evalua.n2
#     campo2.save()
#     campo3 = ma.campo('N3')
#     campo3.valor = evalua.n3
#     campo3.save()
#     campo4 = ma.campo('RECUP')
#     campo4.valor = evalua.recuperacion
#     campo4.save()
#     print ma.id



# for materiaasignada in MateriaAsignada.objects.all().order_by('-id'):
#     evaluacion = materiaasignada.evaluacion()
#     codigo = None
#     datos = {"result": "ok"}
#     if modeloevaluativo_generico():
#         modeloevaluativo = materiaasignada.materia.modeloevaluativo
#         exec modeloevaluativo.logicamodelo
#         calculo_modelo_evaluativo(materiaasignada)
#         materiaasignada.notafinal = round(materiaasignada.notafinal, modeloevaluativo.notafinaldecimales)
#         materiaasignada.save()
#         camposdependientes = []
#         encurso = True
#         for campomodelo in modeloevaluativo.campos():
#             if campomodelo.actualizaestado and materiaasignada.valor_nombre_campo(campomodelo.nombre) > 0:
#                 encurso = False
#         if not encurso:
#             materiaasignada.actualiza_estado()
#         else:
#             materiaasignada.estado_id = NOTA_ESTADO_EN_CURSO
#             materiaasignada.save()
#     print materiaasignada.id


# for histotico in HistoricoRecordAcademico.objects.filter(Q(materiacurso__cerrado=True) | Q(materiaregular__cerrada=True)).distinct():
#     record  = histotico.recordacademico
#     record.save()
#
# linea = 0
# for row in dataReader:
#     maa = MateriaAsignada.objects.filter(matricula__id=int(row[4]), materia__asignatura__id=int(row[5]))[0]
#     maa.fechaasignacion = convertirfecha(row[0])
#     maa.save()
#     print maa.id

# for lecciongrupo in LeccionGrupo.objects.all():
#     # print lecciongrupo.id
#     if lecciongrupo.porciento_asistencia() == 0:
#         # for asistencia in AsistenciaLeccion.objects.filter(leccion__lecciongrupo=lecciongrupo):
#         #     if asistencia.materiaasignada == None:
#         #         print "retirado"
#         #         asistencia.delete()
#         materiasasignadas = MateriaAsignada.objects.filter(id__in=[x.materiaasignada.id for x in AsistenciaLeccion.objects.filter(leccion__lecciongrupo=lecciongrupo).distinct()])
#         todaslecciones = lecciongrupo.lecciones.all()
#         for lecciones in todaslecciones:
#             lecciones.delete()
#         lecciongrupo.delete()
#         for maa in materiasasignadas:
#             maa.save()
#         print "eliminada"




#
# for carrera in Carrera.objects.all():
#     if not GrupoCoordinadorCarrera.objects.filter(group__id=45, carrera=carrera).exists():
#         gc = GrupoCoordinadorCarrera(group__id=45, carrera=carrera)
#         gc.save()
#         print "adicionado"


# for r in RespuestaEvaluacion.objects.filter(materia__isnull=False, evaluador__isnull=True).order_by('-id'):
#     profesor = r.profesor
#     materia = r.materia
#     profesormateria = profesor.profesormateria_set.filter(materia=materia)[0]
#     r.profesormateria = profesormateria
#     r.save()
#     print r.id

# linea = 0
# for row in dataReader:
#     if Inscripcion.objects.filter(persona__cedula=row[0].strip(), carrera_id=int(row[1])).exists():
#         inscripcion = Inscripcion.objects.filter(persona__cedula=row[0].strip(), carrera_id=int(row[1]))[0]
#         inscripcion.fechainicioconvalidacion = convertirfecha(row[2].strip())
#         inscripcion.save()
#     else:
#         print row[0] +', ' +row[1] +', ' + row[2]


# a = []
# for lg in LeccionGrupo.objects.all():
#     if not lg in a and LeccionGrupo.objects.filter(fecha=lg.fecha, horaentrada=lg.horaentrada, horasalida=lg.horasalida, profesor=lg.profesor).exclude(id=lg.id).exists():
#         for lgo in LeccionGrupo.objects.filter(fecha=lg.fecha, horaentrada=lg.horaentrada, horasalida=lg.horasalida, profesor=lg.profesor).exclude(id=lg.id):
#             a.append(lgo)
#
# for i in a:
#     print i
#     lecciones = i.lecciones.all()
#     lecciones.delete()
#     i.delete()


# for ma in MateriaAsignada.objects.all().order_by('-id'):
#     lecciones = ma.materia.lecciones_individuales()
#     asistencias = ma.asistencialeccion_set.all()
#     if asistencias.count() != lecciones.count():
#         ma.asistencialeccion_set.all().delete()
#         ma.asistencias()
#         print "ACTUALIZADOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOO"
#     ma.save()
#     print ma.id


# VERIFICA RECORDS TENGA HISTORICO
# for r in RecordAcademico.objects.annotate(ch=Count('historicorecordacademico__id')).filter(ch=0).order_by('-id'):
#     print r.id
#     r.mi_historico()


# for materia in MateriaAsignada.objects.all().order_by('-id'):
#     if HistoricoRecordAcademico.objects.filter(inscripcion=materia.matricula.inscripcion, asignatura=materia.materia.asignatura, fecha=materia.materia.nivel.fin).exists():
#         hist = HistoricoRecordAcademico.objects.filter(inscripcion=materia.matricula.inscripcion, asignatura=materia.materia.asignatura, fecha=materia.materia.nivel.fin)[0]
#         hist.materiaregular = materia.materia
#         hist.save()
#         print materia.id


# for lg in LeccionGrupo.objects.all():
#     if LeccionGrupo.objects.filter(fecha=lg.fecha, profesor=lg.profesor, horaentrada=lg.horaentrada).exclude(id=lg.id).exists():
#         otras = LeccionGrupo.objects.filter(fecha=lg.fecha, profesor=lg.profesor, horaentrada=lg.horaentrada).exclude(id=lg.id)
#         print str(otras.count()) + ' - ' + str(lg.id)
#         otras.delete()


#LeccionGrupo.objects.filter(horasalida=None).delete()


# for participacion in EvaluacionLeccion.objects.all().order_by('-id'):
#     participacion.materiaasignada = participacion.matricula.materiaasignada_set.filter(materia=participacion.leccion.clase.materia)[0] if participacion.matricula.materiaasignada_set.filter(materia=participacion.leccion.clase.materia).exists() else None
#     participacion.save()
#     print participacion.id


# for ma in MateriaAsignada.objects.all().order_by('-id'):
#     eu = ma.evaluacion_isac()
#     if eu.nota_parcial_1():
#         c = micampo(ma, 'N1')
#         c.valor = eu.n1
#         c.save()
#         c = micampo(ma, 'N2')
#         c.valor = eu.n2
#         c.save()
#         c = micampo(ma, 'P1')
#         c.valor = eu.nota_parcial_1()
#         c.save()
#     if eu.nota_parcial_2():
#         c = micampo(ma, 'N3')
#         c.valor = eu.n3
#         c.save()
#         c = micampo(ma, 'N4')
#         c.valor = eu.n4
#         c.save()
#         c = micampo(ma, 'P2')
#         c.valor = eu.nota_parcial_2()
#         c.save()
#     if eu.recuperacion:
#         actualizar = True
#         c = micampo(ma, 'MEJO')
#         c.valor = eu.recuperacion
#         c.save()
#         print 'actualizo'
#     print ma.id






# contador = 0
# for row in dataReader:
#     if DocumentoColeccion.objects.filter(codigo=row[0].strip()).exists():
#         documento = DocumentoColeccion.objects.filter(codigo=row[0].strip())[0]
#         dc = documento.documento.documento_carrera()
#         dc.carrera.add(Carrera.objects.get(pk=int(row[1])))
#     contador += 1
#     print contador


# caja = LugarRecaudacion.objects.get(pk=12)
# print caja.persona
# for x in range(1, 32):
#     fin = date(2015, 3, x)
#     inicio = date(2015, 3, x)
#     valor = Pago.objects.filter(fecha__gte=inicio, fecha__lte=fin, sesion__caja=caja).aggregate(total=Sum('valor'))['total']
#     print x
#     print valor

# fechainicio = date(2015,1,1)

# for pago in Pago.objects.filter(fecha__gte=fechainicio):
#     if pago.sesion.fecha != pago.fecha:
#         factura = pago.factura_set.all()[0]
#         print factura
#         print pago.sesion
#         if SesionCaja.objects.filter(caja=pago.sesion.caja, fecha=factura.fecha).exists():
#            sesion =  SesionCaja.objects.filter(caja=pago.sesion.caja, fecha=factura.fecha)[0]
#            pago.sesion = sesion
#            pago.save()
#            print "cambio"
#         else:
#             print "no existe sesion de caja"
# for idioma in IdiomaDomina.objects.all().order_by('-id'):
#     du = idioma.persona.datos_unificados()
#     idioma.datosunificados = du
#     idioma.save()
#     print idioma.id
# print "TERMINO IDIOMA"
# print "TERMINO IDIOMA"
#
# for estudios in EstudioInscripcion.objects.all().order_by('-id'):
#     du = estudios.inscripcion.persona.datos_unificados()
#     estudios.datosunificados = du
#     estudios.save()
#     print estudios.id
# print "TERMINO ESTUDIOS"
# print "TERMINO ESTUDIOS"
#
# for seguimiento in SeguimientoEstudiante.objects.all().order_by('-id'):
#     du = seguimiento.inscripcion.persona.datos_unificados()
#     seguimiento.datosunificados = du
#     seguimiento.save()
#     print seguimiento.id
# print "TERMINO SEGUIMIENTO"
# print "TERMINO SEGUIMIENTO"
#
# for conocimiento in ConocimientoAdiconal.objects.all().order_by('-id'):
#     du = conocimiento.persona.datos_unificados()
#     conocimiento.datosunificados = du
#     conocimiento.save()
#     print conocimiento.id
# print "TERMINO CONOCIMIENTO ADICIONAL"
# print "TERMINO CONOCIMIENTO ADICIONAL"
#
# for conocimiento in ConocimientoInformatico.objects.all().order_by('-id'):
#     du = conocimiento.inscripcion.persona.datos_unificados()
#     conocimiento.datosunificados = du
#     conocimiento.save()
#     print conocimiento.id
# print "TERMINO CONOCIMIENTO INFORMATICO"
# print "TERMINO CONOCIMIENTO INFORMATICO"
#



# for ficha in FichaSocioEconomica.objects.all().order_by('-id'):
#     du = ficha.inscripcion.persona.datos_unificados()
#     ficha.datosunificados = du
#     ficha.save()
#     print ficha.id
# print "TERMINO FICHA SOCIOECONOMICA"
# print "TERMINO FICHA SOCIOECONOMICA"

# linea = 0
# for row in dataReader:
#     if not Asignatura.objects.filter(nombre=row[0]).exists():
#         asignatura = Asignatura(nombre=row[0],
#                                 codigo='',
#                                 creditos=int(row[1]))
#         asignatura.save()
#         linea += 1
#         print linea
# for especialidad in Especialidad.objects.all():
#     if Especialidad.objects.filter(nombre=especialidad.nombre).exclude(id=especialidad.id).exists():
#         antigua = Especialidad.objects.filter(nombre=especialidad.nombre).exclude(id=especialidad.id)[0]
#         for inscripcion in Inscripcion.objects.filter(especialidad=antigua):
#             inscripcion.especialidad = especialidad
#             inscripcion.save()
#         antigua.delete()


# for inscripcion in Inscripcion.objects.all():
#     grupo = inscripcion.grupo()
#     if grupo:
#         inscripcion.sede = grupo.sede
#         inscripcion.modalidad = grupo.modalidad
#         inscripcion.sesion = grupo.sesion
#         inscripcion.carrera = grupo.carrera
#         inscripcion.save()
#     print inscripcion.id


# ERRORES = 0
#
# def to_unicode(s):
#     if isinstance(s, unicode): return s
#
#     from locale import getpreferredencoding
#     for cp in (getpreferredencoding(), "cp1255", "cp1250"):
#         try:
#             return unicode(s, cp)
#         except UnicodeDecodeError:
#             pass
#     raise Exception("Conversion to unicode failed")
# or fallback like:
# return unicode(s, getpreferredencoding(), "replace")







#Arreglo Record e Historicos - Asignaturas q NO APLICAN por vieja malla segun niveles
# linea = 1
# for row in dataReader:
#     carrera = Carrera.objects.get(pk=int(row[1].strip()))
#     if Inscripcion.objects.filter(persona__cedula=row[0].strip(), carrera=carrera).exists():
#         inscripcion = Inscripcion.objects.filter(persona__cedula=row[0].strip(), carrera=carrera)[0]
#         im = inscripcion.malla_inscripcion()
#         if im:
#             asignaturas_mallas = im.malla.asignaturamalla_set.filter(nivelmalla__id__lte=int(row[2].strip()))
#             for am in asignaturas_mallas:
#                 if not inscripcion.recordacademico_set.filter(asignatura=am.asignatura).exists():
#                     record = RecordAcademico(inscripcion=inscripcion,
#                                              asignatura=am.asignatura,
#                                              nota=0,
#                                              asistencia=0,
#                                              fecha=datetime.now().date(),
#                                              aprobada=True,
#                                              pendiente=False,
#                                              creditos=0,
#                                              horas=0,
#                                              noaplica=True,
#                                              observaciones="NO APLICA - MALLA ANTIGUA",
#                                              homologada=False,
#                                              valida=False)
#                     record.save()
#                     record.actualizar()
#                     matricula = inscripcion.matricula()
#                     if matricula:
#                         if matricula.materiaasignada_set.filter(materia__asignatura=am.asignatura).exists():
#                             maa = matricula.materiaasignada_set.filter(materia__asignatura=am.asignatura)[0]
#                             maa.delete()
#                             print "Elimina Materia de Matricula: " + str(linea)
#                     print "Actualizado: " + str(linea) + " (" + inscripcion.persona.cedula + ")"
#     else:
#         data = [[row[0],row[1],row[2],row[3],row[4], "I"]]
#         dataWriter.writerows(data)
#         print "fallo"
#
#     print linea
#     linea += 1

# DUPLICADOS RECORS
# for inscripcion in Inscripcion.objects.all():
#     for record in inscripcion.recordacademico_set.all():
#         if inscripcion.recordacademico_set.filter(asignatura=record.asignatura).count() > 1:
#             print record.id.__str__()+', '+inscripcion.persona.cedula + ', ' + record.asignatura.nombre + ', ' + record.nota.__str__()
#


# CONFLICTO DE DOCENTES
# for profesor in Profesor.objects.all():
#     materias = profesor.materias_imparte()
#     conflicto = conflicto_materias_seleccionadas(materias)
#     if conflicto:
#         print profesor
#         print conflicto


# # eliminar matricula
# matri = Matricula.objects.filter(nivel__periodo=7, inscripcion__carrera_id=1)
# for matri1 in matri:
#     for materiaasignada in matri1.materiaasignada_set.all():
#         materiaasignada.delete()
#         print 'listo detalle'
#     matri1.delete()
# print 'listo total'


# conflicto de horario (carlos)
# tipo=TipoProfesor.objects.get(pk=2)
# profesor=ProfesorMateria.objects.filter(materia__nivel__periodo_id=7,materia__asignaturamalla__malla__carrera__id__in=[1], materia__asignatura__id__in=[1467,1468,1469,1470])
# for profesor1 in profesor:
#     profesor1.tipoprofesor=tipo
#     profesor1.principal=False
#     profesor1.save()
# print 'listo'



# conflicto de horario (carlos)
salud=[1,3,4]
ingenieria=[22,24]
administrativa=[5,6,7,8,11]
educacion=[13,14,15,16,18,]
distancia=[19,21,26]
#
# clases=Clase.objects.filter(materia__nivel__periodo_id=7,materia__asignaturamalla__malla__carrera__id__in=salud).order_by('materia__asignaturamalla__malla__carrera__id')
# for clase1 in clases:
#     turno = clase1.turno
#     dia = clase1.dia
#     materia = clase1.materia
#     inicio = clase1.inicio
#     fin = clase1.fin
#     profesorprincipal = clase1.materia.profesor_principal()
#     for clase in Clase.objects.filter(Q(activo=True) & Q(materia__profesormateria__principal=True) & Q(materia__profesormateria__profesor=profesorprincipal) & Q(dia=dia) & ((Q(turno__comienza=turno.comienza) & Q(turno__termina=turno.termina)) | (Q(turno__termina__gte=turno.comienza) & Q(turno__termina__lte=turno.termina)) | (Q(turno__comienza__gte=turno.comienza) & Q(turno__comienza__lte=turno.termina)))).exclude(materia_id=materia.id):
#         if (clase.inicio <= inicio <= clase.fin) or (clase.inicio <= fin <= clase.fin) or (clase.inicio >= inicio and clase.fin <= fin):
#              print "En la carrera:" + clase1.materia.asignaturamalla.malla.carrera.nombre_completo()  + " la Materia: " + clase1.materia.asignatura.nombre + clase1.materia.nivel.nombre_corto() + clase1.materia.paralelo + " tiene conflicto de horario en: " + clase.nombre_conflicto_docente()

# materias=Materia.objects.filter(nivel__periodo_id=7,asignaturamalla__malla__carrera__id=1, nivel__nivelmalla__id=4, nivel__sesion__id=1)
# materias=Materia.objects.filter(nivel__periodo__id=7,asignaturamalla__malla__carrera__id=1, asignaturamalla__nivelmalla__id=4, nivel__sesion__id=1, paralelo='A2')
# for clasesmaterias in materias:
#     claseseleccionada = clasesmaterias.id
#     clasesmaterias = Clase.objects.filter(materia__nivel__periodo__id=7,materia__asignaturamalla__malla__carrera__id=1, materia__asignaturamalla__nivelmalla__id=4, materia__nivel__sesion__id=1, materia__paralelo='A2').exclude(id=claseseleccionada)
#     for clase in clasesmaterias:
#         # clasesverificadas = []
#         # clasesverificadas.append(clase.id)
#         # clasesmaterias1=clase.clase
#         if clasesmaterias.filter((Q(turno__comienza__lte=clase.turno.termina, turno__termina__gte=clase.turno.termina,dia=clase.dia) | Q(turno__comienza__lte=clase.turno.comienza,turno__termina__gte=clase.turno.comienza, dia=clase.dia)) & (Q(inicio__lte=clase.inicio, fin__gte=clase.inicio) | Q(inicio__lte=clase.fin, fin__gte=clase.fin))).exclude(materia__id=claseseleccionada).exists():
#             conflicto = clasesmaterias.filter((Q(turno__comienza__lte=clase.turno.termina,turno__termina__gte=clase.turno.termina, dia=clase.dia) | Q(turno__comienza__lte=clase.turno.comienza, turno__termina__gte=clase.turno.comienza, dia=clase.dia)) & (Q(inicio__lte=clase.inicio, fin__gte=clase.inicio) | Q(inicio__lte=clase.fin,fin__gte=clase.fin))).exclude(materia__id=claseseleccionada)[0]
#             print "conflicto de horario: " + to_unicode(clase.materia.asignatura.nombre) + "(" + to_unicode(clase.materia.identificacion) + ") y " + to_unicode(conflicto.materia.asignatura.nombre) + "(" + to_unicode(conflicto.materia.identificacion) + ") DIA: " + conflicto.dia.__str__()
#

# # # conflicto horario docente cruces
def conflicto_materias_seleccionadas(materias):
    texto = ""
    if CHEQUEAR_CONFLICTO_HORARIO:
        # materias = materias.filter(cerrado=False, fin__gte=datetime.now().date())
        clasesmaterias = Clase.objects.filter(materia__in=materias, activo=True, fin__gte=datetime.now().date()).order_by('dia')
        clasesverificadas = []
        materiasverificadas = []
        for clase in clasesmaterias:
            clasesverificadas.append(clase.id)
            materiasverificadas.append(clase.materia.id)
            if clasesmaterias.filter((Q(turno__comienza__lte=clase.turno.termina, turno__termina__gte=clase.turno.termina, dia=clase.dia) | Q(turno__comienza__lte=clase.turno.comienza, turno__termina__gte=clase.turno.comienza, dia=clase.dia)) & (Q(inicio__lte=clase.inicio, fin__gte=clase.inicio) | Q(inicio__lte=clase.fin, fin__gte=clase.fin))).exclude(materia__id__in=materiasverificadas).exists():
                conflicto = clasesmaterias.filter((Q(turno__comienza__lte=clase.turno.termina, turno__termina__gte=clase.turno.termina, dia=clase.dia) | Q(turno__comienza__lte=clase.turno.comienza, turno__termina__gte=clase.turno.comienza, dia=clase.dia)) & (Q(inicio__lte=clase.inicio, fin__gte=clase.inicio) | Q(inicio__lte=clase.fin, fin__gte=clase.fin))).exclude(materia__id__in=materiasverificadas)[0]
                texto = texto + "conflicto de horario: " + to_unicode(clase.materia.asignatura.nombre) + "(" + to_unicode(clase.materia.identificacion) + ") y " + to_unicode(conflicto.materia.asignatura.nombre) + "(" + to_unicode(conflicto.materia.identificacion) + ") DIA: " + conflicto.dia.__str__() +"\n"
    return texto
#
#
# periodo=7
# salud=[1,3,4]
# ingenieria=[22,24]
# administrativa=[5,6,7,8,11]
# educacion=[13,14,15,16,18,]
# distancia=[19,21,26]
# materias=Materia.objects.values_list('asignaturamalla__malla__carrera_id','nivel__sesion__id','asignaturamalla__nivelmalla__id','paralelo')._next_is_sticky().filter(nivel__periodo__id=periodo,asignaturamalla__malla__carrera__id__in=salud).order_by('nivel__sesion__id','asignaturamalla__malla__carrera__id','asignaturamalla__nivelmalla__id').distinct()
# data=[]
# for materiaselecciona in materias:
#     carreraid=materiaselecciona[0]
#     nivelid=materiaselecciona[2]
#     sesionid=materiaselecciona[1]
#     paralelonombre=materiaselecciona[3]
#     materia_verifica=Materia.objects.filter(asignaturamalla__malla__carrera_id=carreraid, nivel__sesion__id=sesionid, asignaturamalla__nivelmalla__id=nivelid, paralelo=paralelonombre)
#     print 'Carrera: %s, seccion: %s, nivel: %s, paralelo: %s, %s' % (carreraid, sesionid, nivelid, paralelonombre,conflicto_materias_seleccionadas(materia_verifica))
# print 'listo'


# alumnos por problemas de asistencia al agregarle y despues quitarle
# for materia in MateriaAsignada.objects.filter(Q(evaluaciongenerica__detallemodeloevaluativo__nombre='EX2',evaluaciongenerica__valor__gt=0) | Q(evaluaciongenerica__detallemodeloevaluativo__nombre='EX',evaluaciongenerica__valor__gt=0)).filter(matricula__nivel__periodo__id=6, estado__id=2, asistenciafinal__lt=70, sinasistencia=False).distinct():
#     print u'Estudiante: %s, en la materia: %s, tiene porcentaje de asistencia: %s, tiene nota final: %s' % (materia.matricula.inscripcion.persona.nombre_completo(),materia.materia.asignatura.nombre,materia.asistenciafinal,materia.notafinal)
#
# conflicto de horario este vale
# periodo=Periodo.objects.get(pk=9)
# profesor=Profesor.objects.filter(profesormateria__materia__nivel__periodo=periodo,profesormateria__principal=True).distinct().exclude(persona__apellido1='FACULTAD')
# conflicto=[]
# for pm in profesor:
#     result = conflicto_materias_seleccionadas(pm.materias_imparte_periodo(periodo))
#     if result.__len__():
#         conflicto.extend(result)
#
# for c in conflicto:
#     print c

# for pm in profesor:
#     profesormateria=pm.mis_materias(periodo)
#     for profesormateria1 in profesormateria:
#         bandera=True
#         clase1 = Clase.objects.filter(materia=profesormateria1.materia)
#         for clase in clase1:
#             turno = clase.turno
#             dia = clase.dia
#             inicio = profesormateria1.materia.inicio
#             fin = profesormateria1.materia.fin
#             clases = Clase.objects.exclude(id=clase.id).filter(Q(activo=True) & Q(materia__profesormateria__principal=True) & Q(materia__profesormateria__profesor=pm) & Q(dia=dia) & ((Q(turno__comienza=turno.comienza) & Q(turno__termina=turno.termina)) | (Q(turno__termina__gte=turno.comienza) & Q(turno__termina__lte=turno.termina)) | (Q(turno__comienza__gte=turno.comienza) & Q(turno__comienza__lte=turno.termina))))
#             for claseverificaion in clases:
#                 if (claseverificaion.inicio <= inicio <= claseverificaion.fin) or (claseverificaion.inicio <= fin <= claseverificaion.fin) or (claseverificaion.inicio >= inicio and claseverificaion.fin <= fin):
#                     # if claseverificaion.materia not in conflicto:
#                         # conflicto.append(claseverificaion.materia)
#                     mensaje1= u"El docente: %s; con la materia: %s; en la Carrera: %s; el dia: %s; en la hora: %s; Conflicto de horario en: %s " % (pm, profesormateria1.materia.asignatura.nombre, profesormateria1.materia.asignaturamalla.malla.carrera.nombre_completo(),dia,turno ,claseverificaion.nombre_conflicto_docente())
#                     print mensaje1
#                     mensaje=mensaje1.encode('latin-1')
#                     data=[mensaje]
#                     dataWriter.writerows(data)
#
# print 'listo'

# rol = RolPago.objects.all()
# n = 1
# for rol1 in rol:
#     print n
#     n += 1
#     print rol1.persona
#     if DistributivoPersona.objects.filter(persona=rol1.persona).exists():
#         distributivo = DistributivoPersona.objects.filter(estadopuesto__prioridad__gt=0, persona=rol1.persona).order_by('estadopuesto__prioridad').distinct()[0]
#         rol1.denominacionpuesto = distributivo.denominacionpuesto
#         rol1.unidadorganica = distributivo.unidadorganica
#         rol1.grado = distributivo.grado
#         rol1.save()
# print 'listo'

# # # modificar fechas nivel, materia y horario
# periodo=Periodo.objects.get(pk=7)
# materia=Materia.objects.filter(asignaturamalla__malla__carrera__in=[1,3,4], inicio=convertirfecha('02-05-2016')).order_by('asignaturamalla__malla__carrera')
# for mate in materia:
#     print " id=%s, Carrera:%s, Materia:%s, Nivel: %s, Fecha Inicio: %s, Fecha Fin: %s" % (mate.id,mate.asignaturamalla.malla.carrera.nombre_completo(),mate.asignatura.nombre, mate.nivel, mate.inicio, mate.fin)
#     nivel=Nivel.objects.filter(materia=mate)[0]
#     nivel.inicio=convertirfecha('10-05-2016')
#     # nivel.fin=convertirfecha('29-06-2016')
#     nivel.save()
#     mate.inicio=convertirfecha('10-05-2016')
#     # mate.fin=convertirfecha('02-06-2016')
#     # mate.fechafinasistencias=convertirfecha('07-09-2016')
#     mate.save()
#     clase=Clase.objects.filter(materia=mate)
#     for cla in clase:
#         cla.inicio=convertirfecha('10-05-2016')
#         # cla.fin=convertirfecha('02-06-2016')
#         cla.save()
# periodo = Periodo.objects.filter(pk=14)[0]
# ID_CRITERIO_HORAS_CLASE_REEMPLAZO = variable_valor('CRITERIO_HORAS_CLASE_REEMPLAZO')
# ActividadDetalleDistributivo.objects.filter(criterio__distributivo__periodo=periodo, criterio__criteriodocenciaperiodo__isnull=False,criterio__criteriodocenciaperiodo__criterio__id__in=[CRITERIO_HORAS_CLASE_AYUDANTIA,CRITERIO_HORAS_CLASE_PRACTICA,CRITERIO_HORAS_CLASE_TIEMPO_COMPLETO_ID,CRITERIO_HORAS_CLASE_MEDIO_TIEMPO_ID,ID_CRITERIO_HORAS_CLASE_REEMPLAZO]).update(horas=0)
# DetalleDistributivo.objects.filter(distributivo__periodo=periodo, criteriodocenciaperiodo__isnull=False,criteriodocenciaperiodo__criterio__id__in=[CRITERIO_HORAS_CLASE_AYUDANTIA,CRITERIO_HORAS_CLASE_PRACTICA,CRITERIO_HORAS_CLASE_TIEMPO_COMPLETO_ID,CRITERIO_HORAS_CLASE_MEDIO_TIEMPO_ID,ID_CRITERIO_HORAS_CLASE_REEMPLAZO]).update(horas=0)
# profesormateria = ProfesorMateria.objects.values_list('profesor','tipoprofesor').filter(materia__nivel__periodo=periodo).distinct()
# for profe1 in profesormateria:
#     print(profe1)
#     idtipoprofesor = profe1[1]
#     profesor = Profesor.objects.filter(pk= profe1[0])[0]
#     # profe1.hora=profe1.materia.horassemanales
#     # profe1.save()
#     profesor.actualizar_distributivo_horas(periodo, idtipoprofesor)
# for detalle in DetalleDistributivo.objects.filter(distributivo__periodo=periodo):
#     print(detalle)
#     detalle.actualiza_padre()
# for detale in DetalleDistributivo.objects.filter(horas=0, distributivo__periodo=periodo):
#     print(detale)
#     detale.delete()
# print('listo')

# periodo=7
# carrera=5
# materias=Materia.objects.filter(nivel__periodo__id=periodo,asignaturamalla__malla__carrera__id=carrera).order_by('nivel__sesion__id','asignaturamalla__malla__carrera__id','asignaturamalla__nivelmalla__id')
# for mate in materias:
#     carreranombre=mate.asignaturamalla.malla.carrera.nombre_completo()
#     nivelid=materiaselecciona.asignaturamalla.nivelmalla.id
#     nivelnombre=materiaselecciona.asignaturamalla.nivelmalla.nombre
#     sesionid=materiaselecciona.nivel.sesion.id
#     sesionnombre=materiaselecciona.nivel.sesion.nombre
#     asignaturanombre=materiaselecciona.asignatura.nombre
#     paralelonombre=materiaselecciona.paralelo
#     materiasele=materiaselecciona.id




# materias sin profesor principal
# materias=Materia.objects.filter(nivel__periodo_id=7).order_by('asignaturamalla__malla__carrera')
# lista = []
# for materia in materias:
#     if not materia.profesormateria_set.filter(tipoprofesor=1).exists():
#         lista.append(materia)
#
# for l in lista:
#     print 'Carrera: %s  - Materia: %s   -  Nivel: %s' % (l.asignaturamalla.malla.carrera.nombre_completo(), l.asignatura.nombre, l.nivel.nombre_corto())




#ARREGLO DE INGLES 1 y BASICO
# inglesbasico = Asignatura.objects.get(pk=782)
# ingles1 = Asignatura.objects.get(pk=783)
# for inscripcion in Inscripcion.objects.all():
#     if inscripcion.recordacademico_set.filter(asignatura=ingles1).exists():
#         record1 = inscripcion.recordacademico_set.filter(asignatura=ingles1)[0]
#         if record1.aprobada:
#             if not inscripcion.recordacademico_set.filter(asignatura=inglesbasico).exists():
#                 record = RecordAcademico(inscripcion=inscripcion,
#                                          asignatura=inglesbasico,
#                                          nota=0,
#                                          asistencia=70,
#                                          fecha=datetime.now().date(),
#                                          aprobada=True,
#                                          pendiente=False,
#                                          creditos=0,
#                                          horas=0,
#                                          noaplica = True,
#                                          observaciones=" (Migracion)",
#                                          homologada=True,
#                                          valida=True)
#                 record.save()
#                 record.actualizar()
#                 print "Actualizado: (" + inscripcion.persona.cedula + ")"
#
#     print inscripcion.id

#Modulos de Ingles - Notas Historicas
# linea = 1
# for row in dataReader:
#     asignatura = Asignatura.objects.get(pk=int(row[3].strip()))
#     carrera = Carrera.objects.get(pk=int(row[1].strip()))
#
#     if Inscripcion.objects.filter(persona__cedula=row[0].strip(), carrera=carrera).exists():
#         inscripcion = Inscripcion.objects.filter(persona__cedula=row[0].strip(), carrera=carrera)[0]
#         if inscripcion.historicorecordacademico_set.filter(asignatura=asignatura).exists():
#             record = inscripcion.historicorecordacademico_set.filter(asignatura=asignatura).order_by('-fecha')[0]
#             record.nota = Decimal(row[4].strip()), 2)
#             record.fecha = convertirfecha(row[6].strip())
#             record.asistencia = 70
#             record.aprobada = True if row[5].strip() == "A" else False
#             record.pendiente = False
#             record.creditos = 3.75
#             record.horas = 70
#             record.observaciones = row[2].strip()[:90] + " (M)"
#             record.homologada = False
#             record.valida = True if row[5].strip() == "A" else False
#             record.noaplica = False
#             record.save()
#             print "Actualizado: " + str(linea) + " (" + inscripcion.persona.cedula + ")"
#         else:
#             record = RecordAcademico(inscripcion=inscripcion,
#                                      asignatura=asignatura,
#                                      nota=Decimal(row[4].strip()), 2),
#                                      asistencia=70,
#                                      fecha=convertirfecha(row[6].strip()),
#                                      aprobada=True if row[5].strip() == "A" else False,
#                                      pendiente=False,
#                                      creditos=3.75,
#                                      horas=70,
#                                      observaciones=row[2].strip()[:90] + " (M)",
#                                      homologada=False,
#                                      noaplica = False,
#                                      valida=True if row[5].strip() == "A" else False)
#             record.save()
#             print "Creado: " + str(linea) + " (" + inscripcion.persona.cedula + ")"
#         record.actualizar()
#
#     else:
#         data = [[row[0],row[1],row[2],row[3],row[4], row[5],"R"]]
#         dataWriter.writerows(data)
#         print "fallo"
#
#     linea += 1



# for record in RecordAcademico.objects.filter(inscripcion__carrera__id=4, asignatura__id=1061):
#     for historico in record.historicorecordacademico_set.all():
#         historico.asignatura_id = 100
#         historico.creditos = 5 if historico.aprobada else 0
#         historico.horas = 80 if historico.aprobada else 0
#         historico.homologada = True
#         historico.save()
#     record.asignatura_id = 100
#     record.creditos = 5 if record.aprobada else 0
#     record.horas = 80 if record.aprobada else 0
#     record.homologada = True
#     record.observaciones = 'MORFOFISIOLOGIA(MIGRACION)'
#     record.save()
#     nueva = RecordAcademico(inscripcion=record.inscripcion,
#                             asignatura_id=622,
#                             nota=record.nota,
#                             asistencia=record.asistencia,
#                             fecha=record.fecha,
#                             noaplica=record.noaplica,
#                             convalidacion=record.convalidacion,
#                             homologada=record.homologada,
#                             aprobada=record.aprobada,
#                             pendiente=record.pendiente,
#                             creditos=5,
#                             horas=80,
#                             valida=record.valida,
#                             observaciones=record.observaciones)
#     nueva.save()
#     nueva.actualizar()
#     print record.id

# linea = 0
# for row in dataReader:
#     if Inscripcion.objects.filter(persona__cedula=row[0], carrera__id=int(row[1])).exists():
#         inscripcion = Inscripcion.objects.filter(persona__cedula=row[0], carrera__id=int(row[1]))[0]
#         if inscripcion.recordacademico_set.filter(asignatura_id=int(row[2])).exists():
#             record = inscripcion.recordacademico_set.filter(asignatura_id=int(row[2]))[0]
#             if record.historicorecordacademico_set.filter(fecha=convertirfecha2(row[4])).exists():
#                 historico = record.historicorecordacademico_set.filter(fecha=convertirfecha2(row[4]))[0]
#                 historico.nota = Decimal(row[3]), 2)
#                 historico.asistencia=70
#                 historico.aprobada=True if row[5][0] == 'A' else False
#                 historico.pendiente=False
#                 historico.creditos=0
#                 historico.convalidacion=False
#                 historico.homologada=False
#                 historico.valida=True if row[5][0] == 'A' else False
#                 historico.observaciones=''
#                 historico.save()
#             else:
#                 historico = HistoricoRecordAcademico(recordacademico=record,
#                                                      inscripcion=inscripcion,
#                                                      asignatura_id=int(row[2]),
#                                                      nota=Decimal(row[3]), 2),
#                                                      asistencia=70,
#                                                      fecha=convertirfecha2(row[4]),
#                                                      convalidacion=False,
#                                                      aprobada=True if row[5][0] == 'A' else False,
#                                                      pendiente=False,
#                                                      creditos=0,
#                                                      homologada=False,
#                                                      valida=True if row[5][0] == 'A' else False,
#                                                      observaciones='')
#                 historico.save()
#             historico.actualizar()
#         else:
#             record = RecordAcademico(inscripcion=inscripcion,
#                                      asignatura_id=int(row[2]),
#                                      nota=Decimal(row[3]), 2),
#                                      asistencia=70,
#                                      fecha=convertirfecha2(row[4]),
#                                      convalidacion=False,
#                                      aprobada=True if row[5][0] == 'A' else False,
#                                      pendiente=False,
#                                      creditos=0,
#                                      homologada=False,
#                                      valida=True if row[5][0] == 'A' else False,
#                                      observaciones='')
#             record.save()
#             record.actualizar()
#     else:
#         data = [[row[0],row[1],row[2],row[3],row[4],row[5],row[6], "I"]]
#         dataWriter.writerows(data)
#         print "fallo"
#     linea += 1
#     print linea


# linea = 1
# carrera = Carrera.objects.get(pk=26)
# for row in dataReader:
#     asignatura = Asignatura.objects.get(pk=int(row[1].strip()))
#     fecha = convertirfecha(row[3].strip())
#     if Inscripcion.objects.filter(persona__cedula=row[0].strip(), carrera=carrera).exists():
#         inscripcion = Inscripcion.objects.filter(persona__cedula=row[0].strip(), carrera=carrera)[0]
#         if inscripcion.recordacademico_set.filter(asignatura=asignatura).exists():
#             record = inscripcion.recordacademico_set.filter(asignatura=asignatura)[0]
#             if record.historicorecordacademico_set.filter(fecha=fecha).exists():
#                 historico = record.historicorecordacademico_set.filter(fecha=fecha)[0]
#                 historico.nota = Decimal(row[2].strip()), 2)
#                 historico.asistencia=70
#                 historico.aprobada=True if row[4].strip() == "A" else False
#                 historico.pendiente=False
#                 historico.creditos=0
#                 historico.horas=0
#                 historico.convalidacion=False
#                 historico.homologada=False
#                 historico.valida=True
#                 historico.observaciones="Migracion (05-07-15)"
#                 historico.save()
#             else:
#                 historico = HistoricoRecordAcademico(recordacademico=record,
#                                                      inscripcion=inscripcion,
#                                                      asignatura=asignatura,
#                                                      nota=Decimal(row[2].strip()), 2),
#                                                      asistencia=70,
#                                                      fecha=fecha,
#                                                      convalidacion=False,
#                                                      aprobada=True if row[4].strip() == "A" else False,
#                                                      pendiente=False,
#                                                      creditos=0,
#                                                      horas=0,
#                                                      homologada=False,
#                                                      valida=True,
#                                                      observaciones="Migracion (07-05-15)")
#                 historico.save()
#             historico.actualizar()
#             print "Historico: " + str(record.id) + " (" + str(linea) + ")"
#         else:
#             record = RecordAcademico(inscripcion=inscripcion,
#                                      asignatura=asignatura,
#                                      nota=Decimal(row[2].strip()), 2),
#                                      asistencia=70,
#                                      fecha=fecha,
#                                      convalidacion=False,
#                                      aprobada=True if row[4].strip() == "A" else False,
#                                      pendiente=False,
#                                      creditos=0,
#                                      horas=0,
#                                      homologada=False,
#                                      valida=True,
#                                      observaciones="Migracion (07-05-15)")
#             record.save()
#             record.actualizar()
#             print "Record: " + str(record.id) + " (" + str(linea) + ")"
#     else:
#         data = [[row[0],row[1],row[2],row[3],row[4], "I"]]
#         dataWriter.writerows(data)
#         print "fallo"
#
#     linea += 1
# print linea


# periodo = Periodo.objects.get(pk=6)
# linea = 0
# for matricula in Matricula.objects.filter(nivel__periodo=periodo):
#     materias_todas = matricula.materiaasignada_set.all()
#     materias_pendientes = matricula.materiaasignada_set.filter(matriculas__gt=1)
#     if materias_pendientes.exists():
#         if materias_todas.count() == materias_pendientes.count():
#             print matricula



# for inscripcion in Inscripcion.objects.all():
#     for record in inscripcion.recordacademico_set.all():
#         if record.historicorecordacademico_set.count() >= 4:
#             print inscripcion.persona.nombre_completo() + " - " + record.asignatura.nombre + " - " + record.historicorecordacademico_set.count().__str__()



# lista = []
# for inscripcion in Inscripcion.objects.all():
#     cedula = inscripcion.persona.cedula
#     if Inscripcion.objects.filter(persona__cedula=cedula).count() > 1:
#         if inscripcion not in lista:
#             lista.append(inscripcion)
#


# for l in lista:
#     print l.persona.nombre_completo() + "," + l.carrera.__unicode__() + "," + str(l.id)
#     if inscripcion.recordacademico_set.filter(asignatura_id=605).exists():
#         record = inscripcion.recordacademico_set.filter(asignatura_id=605)[0]
#         if not inscripcion.recordacademico_set.filter(asignatura_id=695).exists():
#             print inscripcion
#             his = HistoricoRecordAcademico(inscripcion=record.inscripcion,
#                                            asignatura_id=695,
#                                            nota=record.nota,
#                                            asistencia=record.asistencia,
#                                            fecha=record.fecha,
#                                            convalidacion=record.convalidacion,
#                                            aprobada=record.aprobada,
#                                            pendiente=record.pendiente,
#                                            creditos=record.creditos,
#                                            homologada=record.homologada,
#                                            valida=record.valida,
#                                            observaciones='')
#             his.save()
#             his.actualizar()
#             inscripcion.actualizar_creditos()

# if inscripcion.recordacademico_set.filter(asignatura_id=605).exists():
#     if inscripcion.recordacademico_set.filter(asignatura_id=695).exists():
#         print "Falta I"



# for clase in Clase.objects.all():
#     clase.inicio = clase.materia.inicio
#     clase.fin = clase.materia.fin
#     clase.save()
#     print clase.id

# for inscripcion in Inscripcion.objects.all():
#     if not inscripcion.tipo_inscripcion():
#         tipo = InscripcionTipoInscripcion(inscripcion=inscripcion, tipoinscripcion_id=1)
#         tipo.save()
#         print inscripcion.id


# for inscripcion in Inscripcion.objects.all().order_by('id'):
#     documentos = inscripcion.documentos_entregados()
#     documentos.pre = True
#     documentos.save()
#     inscripcion.actualizar_nivel()
#     # inscripcion.actualizar_creditos()
#     print inscripcion.id

# for persona in Persona.objects.all():
#     if persona.usuario.groups.filter(id=ALUMNOS_GROUP_ID).exists():
#         if persona.usuario.groups.count() == 1:
#             inscripcion = persona.inscripcion()
#             if not inscripcion:
#                 persona.delete()
#                 print 'D'
#         else:
#             print persona.id


# for malla in Malla.objects.all():
#     for asig in malla.asignaturamalla_set.all():
#         asignatura = asig.asignatura
#         for precedencia in asignatura.precedencia.all():
#             if malla.asignaturamalla_set.filter(asignatura=precedencia, nivelmalla_id__lt=asig.nivelmalla.id).exists():
#                 precedencia= malla.asignaturamalla_set.filter(asignatura=precedencia, nivelmalla_id__lt=asig.nivelmalla.id)[0]
#                 asigmallaprecedencia = AsignaturaMallaPredecesora(asignaturamalla=asig,
#                                                                   predecesora=precedencia)
#                 asigmallaprecedencia.save()
#
#

# linea = 0
# for row in dataReader:
#     if not Documento.objects.filter(codigo=row[0]).exists():
#         documento = Documento(codigo=row[0],
#                               nombre=to_unicode(row[1]),
#                               edicion=to_unicode(row[2])[:39],
#                               anno=int(row[3]),
#                               autor=to_unicode(row[4])[:199],
#                               ubicacionfisica=row[5],
#                               tipo_id=2,
#                               fisico=True,
#                               copias=1,
#                               sede_id= 1,
#                               idioma_id=2)
#         documento.save()
#         linea += 1
#         print linea
#     else:
#         print "Fallo"


# linea = 0
# for row in dataReader:
#     if not Documento.objects.filter(codigo=row[0]).exists():
#         documento = Documento(codigo=row[0],
#                               ubicacionfisica=row[1],
#                               nombre=to_unicode(row[2])[:399],
#                               autor=to_unicode(row[3])[:199],
#                               anno=int(row[4]),
#                               tipo_id=4,
#                               fisico=True,
#                               digital='',
#                               copias=1,
#                               sede_id= 1,
#                               idioma_id=1)
#         documento.save()
#         linea += 1
#         print linea
#     else:
#         print "Fallo"


# linea = 0
# for row in dataReader:
#     if Inscripcion.objects.filter(persona__cedula=row[0], carrera__id=int(row[1])).exists():
#         inscripcion = Inscripcion.objects.filter(persona__cedula=row[0], carrera__id=int(row[1]))[0]
#         if not HistoricoRecordAcademico.objects.filter(inscripcion=inscripcion, asignatura_id=int(row[3]), fecha=convertirfecha(row[5])).exists():
#             his = HistoricoRecordAcademico(inscripcion=inscripcion,
#                                            asignatura_id=int(row[3]),
#                                            nota=Decimal(row[4]), 2),
#                                            asistencia=70,
#                                            fecha=convertirfecha(row[5]),
#                                            convalidacion=False,
#                                            aprobada=True if row[6] == 'A' else False,
#                                            pendiente=False,
#                                            creditos=0,
#                                            homologada=False,
#                                            valida=True if row[6] == 'A' else False,
#                                            observaciones=row[2])
#             his.save()
#             his.actualizar()
#         else:
#             data = [[row[0],row[1],row[2],row[3],row[4],row[5],row[6], "R"]]
#             dataWriter.writerows(data)
#             print "repetida"
#     else:
#         data = [[row[0],row[1],row[2],row[3],row[4],row[5],row[6], "I"]]
#         dataWriter.writerows(data)
#         print "fallo"
#     linea += 1
#     print linea
#
#
# RecordAcademico.objects.filter().distinct()




# linea = 0
# for inscripcion in Inscripcion.objects.all():
#     # inscripcion.actualizar_creditos()
#     inscripcion.actualizar_nivel()
#     linea += 1
#     print linea



# fecha = date(2011,10,31)
#
# for historico in HistoricoRecordAcademico.objects.filter(fecha=fecha):
#     historico.fecha= date(2011,9,30)
#     historico.save()
#     print historico.id
#
# for record in RecordAcademico.objects.filter(fecha= fecha):
#     record.fecha= date(2011,9,30)
#     record.save()
#     print record.id


# for historico in HistoricoRecordAcademico.objects.all():
#     if historico.recordacademico.historicorecordacademico_set.filter(nota=historico.nota, asistencia=historico.asistencia, aprobada=historico.aprobada).count() > 1:
#         if not historico.tiene_actas():
#             historico.delete()
#             print "delete"
#     print historico.id

# for periodo in Periodo.objects.all():
#     for matricula in Matricula.objects.filter(nivel__periodo=periodo):
#         for materia in matricula.materiaasignada_set.all():
#             if HistoricoRecordAcademico.objects.filter(inscripcion=materia.matricula.inscripcion, asignatura=materia.materia.asignatura, fecha=matricula.nivel.fin).exists():
#                 h = HistoricoRecordAcademico.objects.filter(inscripcion=materia.matricula.inscripcion, asignatura=materia.materia.asignatura, fecha=matricula.nivel.fin)[0]
#                 h.nivel = matricula.nivel
#                 h.save()
#         print matricula.id

# for factura in Factura.objects.all():
#     if factura.ruc.__len__() == 10:
#         factura.tipo = 1
#     elif factura.ruc.__len__() == 13:
#         factura.tipo = 2
#     else:
#         factura.tipo = 3
#     factura.save()
#     print factura.id
#
# for cliente in ClienteFactura.objects.all():
#     if cliente.ruc.__len__() == 10:
#         cliente.tipo = 1
#     elif cliente.ruc.__len__() == 13:
#         cliente.tipo = 2
#     else:
#         cliente.tipo = 3
#     cliente.save()
#     print cliente.id


# for actividad in ActividadExtraCurricular.objects.all():
#     if actividad.participanteactividadextracurricular_set.filter(nota__gt=0).exists():
#         actividad.calificar = True
#         actividad.calfmaxima = 100
#         actividad.calfminima = 75
#         actividad.asistminima = 0
#         actividad.save()
#     else:
#         actividad.calificar = False
#     actividad.save()
#     for participante in actividad.participanteactividadextracurricular_set.all():
#         if participante.asistencia == None:
#             participante.asistencia = 0
#         if participante.nota == None:
#             participante.nota = 0
#         participante.save()




# periodo = Periodo.objects.get(pk=5)
# matriculas = Matricula.objects.filter(nivel__periodo=periodo)
# for matricula in matriculas:
#     matricula.inscripcion.actualizar_creditos()
#     print matricula.inscripcion.mi_nivel().nivel.nombre + "; " + matricula.nivel.paralelo + "; " + remover_caracteres_especiales(matricula.inscripcion.persona.nombre_completo()) + "; " + str(matricula.inscripcion.creditos()) + "; " + str(matricula.inscripcion.recordacademico_set.filter(valida=True).count()) + "; " + str(matricula.inscripcion.promedio_record())


# for inscripcion in Inscripcion.objects.all():
#     if Inscripcion.objects.filter(persona__nombres=inscripcion.persona.nombres, persona__apellido1=inscripcion.persona.apellido1, persona__apellido2=inscripcion.persona.apellido2).exclude(id=inscripcion.id).exists():
#         if inscripcion.datos_unificados():
#             otrasinscripciones = Inscripcion.objects.filter(persona__nombres=inscripcion.persona.nombres, persona__apellido1=inscripcion.persona.apellido1, persona__apellido2=inscripcion.persona.apellido2).exclude(id=inscripcion.id)
#             datos = inscripcion.datos_unificados()
#             for otrainscripcion in otrasinscripciones:
#                 datos.inscripciones.add(otrainscripcion)
#         else:
#             otrainscripcion = Inscripcion.objects.filter(persona__nombres=inscripcion.persona.nombres, persona__apellido1=inscripcion.persona.apellido1, persona__apellido2=inscripcion.persona.apellido2).exclude(id=inscripcion.id)[0]
#             datos = DatosUnificados(cedula=inscripcion.persona.cedula)
#             datos.save()
#             datos.inscripciones.add(inscripcion)
#             datos.inscripciones.add(otrainscripcion)

# for inscripcion in Inscripcion.objects.all():
#     if Inscripcion.objects.filter(persona__cedula=inscripcion.persona.cedula).exclude(id=inscripcion.id).exists():
#         if inscripcion.datos_unificados():
#             otrasinscripciones = Inscripcion.objects.filter(persona__cedula=inscripcion.persona.cedula).exclude(id=inscripcion.id)
#             datos = inscripcion.datos_unificados()
#             for otrainscripcion in otrasinscripciones:
#                 datos.inscripciones.add(otrainscripcion)
#         else:
#             otrainscripcion = Inscripcion.objects.filter(persona__cedula=inscripcion.persona.cedula).exclude(id=inscripcion.id)[0]
#             datos = DatosUnificados(cedula=inscripcion.persona.cedula)
#             datos.save()
#             datos.inscripciones.add(inscripcion)
#             datos.inscripciones.add(otrainscripcion)

# periodo = Periodo.objects.get(pk=21)
# for nivel in periodo.nivel_set.all():
#     for matricula in nivel.matricula_set.all():
#         for rubrocuota in matricula.rubrocuota_set.all():
#             rubro = rubrocuota.rubro
#             rubrootro = RubroOtro(rubro=rubro,
#                                   tipo_id=4,
#                                   descripcion=rubrocuota.nombre_corto())
#             rubrootro.save()
#             rubrocuota.delete()
#         for rubromatricula in matricula.rubromatricula_set.all():
#             rubro = rubromatricula.rubro
#             rubrootro = RubroOtro(rubro=rubro,
#                                   tipo_id=4,
#                                   descripcion=rubromatricula.nombre_corto())
#             rubrootro.save()
#             rubromatricula.delete()
#         for rubrootromatricula in matricula.rubrootromatricula_set.all():
#             rubro = rubrootromatricula.rubro
#             rubrootro = RubroOtro(rubro=rubro,
#                                   tipo_id=4,
#                                   descripcion=rubrootromatricula.nombre_corto())
#             rubrootro.save()
#             rubrootromatricula.delete()


# materiaasignada =  MateriaAsignada.objects.all()
# for materia in materiaasignada:
#     materia.matriculas = materia.matricula.inscripcion.historicorecordacademico_set.filter(asignatura=materia.materia.asignatura, fecha__lt=materia.materia.nivel.fin).count() + 1
#     materia.save()
#     if AgregacionEliminacionMaterias.objects.filter(matricula=materia.matricula, asignatura=materia.materia.asignatura).exists():
#         age = AgregacionEliminacionMaterias.objects.filter(matricula=materia.matricula, asignatura=materia.materia.asignatura)[0]
#         age.matriculas = materia.matriculas
#         age.save()



# agre = AgregacionEliminacionMaterias.objects.all()
# for ag in agre:
#     if ag.agregacion:
#         if ag.matricula.materiaasignada_set.filter(materia__asignatura=ag.asignatura).exists():
#             ma = ag.matricula.materiaasignada_set.filter(materia__asignatura=ag.asignatura)[0]
#             ag.nivelmalla = ma.materia.nivel.nivelmalla
#             ag.creditos = ma.materia.creditos
#             ag.save()
#     else:
#         malla = ag.matricula.inscripcion.malla_inscripcion()
#         if malla.malla.asignaturamalla_set.filter(asignatura=ag.asignatura).exists():
#             am = malla.malla.asignaturamalla_set.filter(asignatura=ag.asignatura)[0]
#             ag.nivelmalla = am.nivelmalla
#             ag.creditos = am.creditos
#             ag.save()


# ins = Inscripcion.objects.all()
# for i in ins:
#     gp = i.grupo()
#     i.modalidad = gp.modalidad
#     i.sede = gp.sede
#     i.sesion = gp.sesion
#     i.carrera = gp.carrera
#     i.save()

# nc = NotaCredito.objects.all()
# for n in nc:
#     if n.saldo > 0:
#         print "saldo"
#         if not n.origennotacredito_set.exists():
#             print "modificada"
#             n.valorinicial = n.total_pagado()
#             n.save()


# matriculados = Matricula.objects.filter(nivel__carrera_id=9)
# for matricula in matriculados:
#     print matricula
#     for rubrocuota in matricula.rubrocuota_set.all():
#         ru = rubrocuota.rubro
#         if matricula.nivel.nivelmalla.id==1:
#             ru.valor = 90
#         else:
#             ru.valor = 90
#         ru.save()



# matriculados = Matricula.objects.all()
# for matricula in matriculados:
#     print matricula
#     matricula.eliminar_rubrosmatricula()
#     nivel = matricula.nivel
#     for pago in nivel.pagonivel_set.all():
#         rubro = Rubro(fecha=datetime.now().date(),
#                       valor=pago.valor,
#                       inscripcion=matricula.inscripcion,
#                       cancelado=False,
#                       fechavence=pago.fecha)
#         rubro.save()
#         if pago.tipo == 0:
#             # MATRICULA
#             rm = RubroMatricula(rubro=rubro,
#                                 matricula=matricula)
#             rm.save()
#         elif 1 <= pago.tipo <= 10:
#             # CUOTA MENSUAL
#             rc = RubroCuota(rubro=rubro,
#                             matricula=matricula,
#                             cuota=pago.tipo)
#             rc.save()
#         else:
#             nombre = TIPOS_PAGO_NIVEL[pago.tipo][1]
#             ro = RubroOtroMatricula(rubro=rubro,
#                                     matricula=matricula,
#                                     descripcion=nombre)
#             ro.save()


# niveles = Nivel.objects.filter(carrera_id=10)
# for nivel in niveles:
#     pagos = nivel.pagonivel_set.filter(tipo__gte=1, tipo__lte=10)
#     for pago in pagos:
#         if nivel.nivelmalla.id == 1:
#             pago.valor = 75
#         else:
#             pago.valor = 70
#         pago.save()


# grupo_profesores = Group.objects.get(pk=PROFESORES_GROUP_ID)
# grupo_estudiantes = Group.objects.get(pk=ALUMNOS_GROUP_ID)
# profesores = Profesor.objects.all()
# for profesor in profesores:
#     if profesor.persona.en_grupo(ALUMNOS_GROUP_ID):
#         usuario = profesor.persona.usuario
#         grupo_estudiantes.user_set.remove(usuario)
#         grupo_profesores.user_set.add(usuario)



#periodo = Periodo.objects.get(pk=3)
#proceso = periodo.proceso_evaluativo()
#instrumentoalumno = proceso.instrumento_alumno()
#instrumentoalumnomateria = proceso.instrumento_alumno_materia()
#

# grupo = Grupo.objects.get(pk=0)
# gruponuevo = Grupo.objects.get(pk=0)
# inscritos = grupo.inscripciongrupo_set.all()
# for inscripcion in inscritos:
#     ig = inscripcion.inscripcion.inscripcion_grupo(gruponuevo)

#matriculas = periodo.matriculados()
#for mat in matriculas:
#    materias = mat.materiaasignada_set.all()
#    for m in materias:
#        for p in m.profesores():
#            if not EvaluacionProfesor.objects.filter(proceso=proceso, instrumento=instrumentoalumnomateria, profesor=p, persona=mat.inscripcion.persona, materia=m.materia).exists():
#                if EvaluacionProfesor.objects.filter(proceso=proceso, instrumento=instrumentoalumno, profesor=p, persona=mat.inscripcion.persona).exists():
#                    ea = EvaluacionProfesor.objects.filter(proceso=proceso, instrumento=instrumentoalumno, profesor=p, persona=mat.inscripcion.persona)[0]
#                    nev = EvaluacionProfesor(proceso=proceso,
#                                             instrumento=instrumentoalumnomateria,
#                                             profesor=p,
#                                             fecha=ea.fecha,
#                                             persona=mat.inscripcion.persona,
#                                             materia=m.materia,
#                                             calificacion=ea.calificacion,
#                                             observaciones="")
#                    nev.save()
#                    datos=ea.datoinstrumentoevaluacion_set.all()
#                    for d in datos:
#                        ndatos = DatoInstrumentoEvaluacion(evaluacion=nev,
#                                                           indicador=d.indicador,
#                                                           valor=d.valor,
#                                                           observaciones=d.observaciones)
#                        ndatos.save()
#                    nev.save()
#
#
#ambitos = instrumentoalumno.ambitoinstrumentoevaluacion_set.all()
#for ambito in ambitos:
#    ambito.instrumento_id = instrumentoalumnomateria.id
#    ambito.save()




#eva = proceso.evaluacionprofesor_set.filter(instrumento=instrumentoalumno)




# for row in dataReader:
#     if Documento.objects.filter(codigo=row[0]).exists():
#         documento = Documento.objects.filter(codigo=row[0])[0]
#         if Persona.objects.filter(cedula=row[2]).exists():
#             persona = Persona.objects.filter(cedula=row[2])[0]
#             prestamo= PrestamoDocumento(documento=documento,
#                                         persona=persona,
#                                         tiempo=24,
#                                         responsableentrega_id=3981,
#                                         fechaentrega=convertirfecha(row[7]),
#                                         horaentrega=datetime(2012, 3, 1, int(row[8][:1]), int(row[8][3:4])).time(),
#                                         entregado=True,
#                                         responsablerecibido_id=3981,
#                                         recibido=True,
#                                         fecharecibido=convertirfecha(row[11]),
#                                         horarecibido=datetime(2012, 3, 1, int(row[12][:1]), int(row[12][3:4])).time())
#             prestamo.save()
#     else:
#         print "Fallo"

# for row in dataReader:
#
#     if Inscripcion.objects.filter(persona__nombres=row[2],persona__apellido1=row[0],persona__apellido2=row[1]).exists():
#         inscripcion= Inscripcion.objects.filter(persona__nombres=row[2],persona__apellido1=row[0],persona__apellido2=row[1])[0]
#         data = [[row[0],row[1],row[2],row[3],inscripcion.id]]
#         print "existe"
#     else:
#         data = [[row[0],row[1],row[2],row[3],""]]
#         print "fallo"
#     dataWriter.writerows(data)

#for row in dataReader:
#    if Inscripcion.objects.filter(persona__cedula__contains=row[4]).exists():
#     inscripcion=Inscripcion.objects.filter(persona__cedula__contains=row[4])[0]
#     rubro = Rubro(fecha=datetime.now().date(),
#                   iva=False,
#                   valor=float(row[3]),
#                   inscripcion=inscripcion,
#                   cancelado=False,
#                   fechavence=datetime.now().date())
#     rubro.save()
#     # tiporubro = TipoOtroRubro.objects.get(pk=4)
#     otro=RubroOtro( rubro = rubro,
#                     tipo_id = 4,
#                     descripcion = "MIGRACION DE DEUDAS")
#     otro.save()




# inscripciones = Inscripcion.objects.all()
# for i in inscripciones:
#     im = i.malla_inscripcion()
#     if i.recordacademico_set.exists():
#         re = i.recordacademico_set.all().order_by('-fecha')[0]
#         if im.malla.inicio > re.fecha:
#             if i.carrera.malla_set.filter(inicio__lte=re.fecha).exists():
#                 ma = i.carrera.malla_set.filter(inicio__lte=re.fecha).order_by('inicio')[0]
#                 im.malla = ma
#                 im.save()



# periodo = Periodo.objects.get(pk=5)
# niveles = Nivel.objects.filter(periodo=periodo)
# for nivel in niveles:
#     mat = nivel.materia_set.all()
#     for m in mat:
#         if AsignaturaMalla.objects.filter(malla=nivel.malla, nivelmalla=nivel.nivelmalla, asignatura=m.asignatura).exists():
#             am = AsignaturaMalla.objects.filter(malla=nivel.malla, nivelmalla=nivel.nivelmalla, asignatura=m.asignatura)[0]
#             if m.creditos <> am.creditos:
#                 print m
#                 m.creditos = am.creditos
#                 m.save()


# # DOCENTES
# linea = 1
# for row in dataReader:
#     persona = Persona(nombres = row[3],
#                       apellido1 = row[1],
#                       apellido2 = row[2],
#                       cedula = row[0],
#                       pasaporte = '',
#                       nacimiento = datetime.now().date(),
#                       sexo_id = 1 if row[4]=="F" else 2,
#                       telefono = "",
#                       telefono_conv = "",
#                       email = "",
#                       emailinst = "")
#     persona.save()
#     persona.cambiar_clave()
#     username = calculate_username(persona)
#     password = DEFAULT_PASSWORD
#     user = User.objects.create_user(username, persona.email, password)
#     user.save()
#     gru = Group.objects.get(pk=1)
#     gru.user_set.add(user)
#     gru.save()
#     persona.usuario = user
#     # persona.emailinst = user.username+'@'+EMAIL_DOMAIN
#     persona.save()
#     profesor = Profesor(persona = persona,
#                         activo = True,
#                         fechaingreso = datetime.now().date(),
#                         dedicacion_id = 1,
#                         categoria_id = 6)
#     profesor.save()
#     print linea
#     linea += 1


# listanio = {}
# historico = HistoricoRecordAcademico.objects.all().order_by('fecha','inscripcion')
# anio = 0
# inscripcion = 0
# nivel1 = 0
# nivel2 = 0
# nivel3 = 0
# nivel4 = 0
# nivel5 = 0
# nivel6 = 0
#
# matriculas_nivel1 = 0
# matriculas_nivel2 = 0
# matriculas_nivel3 = 0
# matriculas_nivel4 = 0
# matriculas_nivel5 = 0
# matriculas_nivel6 = 0
#
# for h in historico:
#     if h.inscripcion.id <> inscripcion:
#         if inscripcion > 0:
#             niveles = [nivel1, nivel2, nivel3, nivel4, nivel5, nivel6]
#             nivel_matricula = niveles.index(max(niveles)) + 1
#             if nivel_matricula == 1:
#                 matriculas_nivel1 += 1
#             elif nivel_matricula == 2:
#                 matriculas_nivel2 += 1
#             elif nivel_matricula == 3:
#                 matriculas_nivel3 += 1
#             elif nivel_matricula == 4:
#                 matriculas_nivel4 += 1
#             elif nivel_matricula == 5:
#                 matriculas_nivel5 += 1
#             elif nivel_matricula == 6:
#                 matriculas_nivel6 += 1
#
#         inscripcion = h.inscripcion.id
#         nivel1 = 0
#         nivel2 = 0
#         nivel3 = 0
#         nivel4 = 0
#         nivel5 = 0
#         nivel6 = 0
#
#     nivel = h.nivel_asignatura()
#     if nivel:
#         id = nivel.id
#         if id == 1:
#             nivel1 += 1
#         elif id == 2:
#             nivel2 += 1
#         elif id == 3:
#             nivel3 += 1
#         elif id == 4:
#             nivel4 += 1
#         elif id == 5:
#             nivel5 += 1
#         elif id == 6:
#             nivel6 += 1
#
#     if h.fecha.year <> anio:
#         matriculas = {"primer": matriculas_nivel1, "segundo" :matriculas_nivel2, "tercero": matriculas_nivel3, "cuarto": matriculas_nivel4,"quinto": matriculas_nivel5, "sexo": matriculas_nivel6}
#         nuevoanio = {"anio_" + anio.__str__(): {"matriculas": matriculas}}
#         listanio.update(nuevoanio)
#         anio = h.fecha.year
#         matriculas_nivel1 = 0
#         matriculas_nivel2 = 0
#         matriculas_nivel3 = 0
#         matriculas_nivel4 = 0
#         matriculas_nivel5 = 0
#         matriculas_nivel6 = 0
#
# print listanio




# linea = 1
# for row in dataReader:
#     codigo = int(row[0]).__str__().zfill(10)
#     if Documento.objects.filter(codigo=codigo).exists():
#         documento = Documento.objects.filter(codigo=codigo)[0]
#         libros = int(row[1])
#         for x in range(1,libros +1,1):
#             if documento.documentocoleccion_set.exists():
#                 codigo = int(documento.documentocoleccion_set.order_by('-id')[0].codigo) + 1
#             else:
#                 codigo = 1
#             ejemplar = DocumentoColeccion(documento=documento,
#                                           codigo=codigo.__str__().zfill(10),
#                                           habilitado=True)
#             ejemplar.save()
#     else:
#         print "Fallo"



#
#  DETALLE DOCUMENTOS BIBLIOTECA
#
# linea = 1
# for row in dataReader:
#     print linea
#     if Documento.objects.filter(codigo=int(row[0])).exists():
#         documento = Documento.objects.filter(codigo=int(row[0]))[0]
#         parametro = int(row[1])
#         if  parametro == 1:
#             documento.edicion = to_unicode(row[6])
#         elif parametro == 2:
#             documento.codigoisbnissn = to_unicode(row[6])
#         elif parametro == 3:
#             documento.paginas = int(row[6])
#         elif parametro == 4:
#             documento.lugarpublicacion = to_unicode(row[6])
#         elif parametro == 5:
#             documento.establecimientoresponsable = to_unicode(row[6])
#         elif parametro == 6:
#             documento.editora = to_unicode(row[6])
#         elif parametro == 7:
#             documento.anno = int(row[6])
#         elif parametro == 8 or parametro == 13:
#             documento.resumen = to_unicode(row[6])
#         elif parametro == 9 or parametro == 10:
#             if documento.descripcionfisica.__len__() <150:
#                 documento.descripcionfisica = documento.descripcionfisica  + " " + to_unicode(row[6])
#         elif parametro == 11:
#             documento.prestamosala = row[6] == 'EN SALA'
#         elif parametro == 12:
#             documento.preciocosto = float(row[6])
#         elif parametro == 20:
#             documento.codigocutter = to_unicode(row[6])
#         elif parametro == 21:
#             documento.codigodewey = to_unicode(row[6])
#         documento.save()
#     linea += 1



#
#  BIBLIOTECA
#
# linea = 1
# for row in dataReader:
#     print linea
#     documento = Documento(codigo=to_unicode(row[0]),
#                          codigoisbnissn='',
#                          nombre=to_unicode(row[6]),
#                          nombre2=to_unicode(row[7]),
#                          ubicacionfisica='',
#                          autor=to_unicode(row[9]),
#                          autorcorporativo='',
#                          lugarpublicacion='',
#                          tipo_id=2,
#                          anno = 1990,
#                          emision='',
#                          edicion='',
#                          palabrasclaves=to_unicode(row[10]),
#                          fisico=True,
#                          copias=1,
#                          paginas=100,
#                          editora='',
#                          sede_id=1,
#                          codigodewey=to_unicode(row[4]),
#                          idioma_id=1,
#                          tutor='',
#                          resumen='',
#                          codigocutter=to_unicode(row[5]),
#                          preciocosto=0,
#                          descripcionfisica='',
#                          establecimientoresponsable=to_unicode(row[8]),
#                          prestamosala=True)
#     documento.save()
#     linea += 1

#
#  RECORD ACADEMICO
#
#identificador = ''
#linea = 1
#for row in dataReader:
#   if Inscripcion.objects.filter(persona__cedula__contains=row[17]).exists():
#       inscr = Inscripcion.objects.filter(persona__cedula__contains=row[17])[0]
#       asignatura = Asignatura.objects.get(pk=int(row[15]))
#       if row[13][0] == 'A':
#            aprobada = True
#       else:
#            aprobada = False
#       if not HistoricoRecordAcademico.objects.filter(inscripcion=inscr, asignatura=asignatura, nota=Decimal(row[12]),2)).exists():
#            his = HistoricoRecordAcademico(inscripcion=inscr,
#                                          asignatura=asignatura,
#                                          nota=Decimal(row[12]),2),
#                                          asistencia=100,
#                                          fecha=convertirfecha(row[18]),
#                                          convalidacion=False,
#                                          aprobada=aprobada,
#                                          pendiente=False,
#                                          creditos=0,
#                                          homologada=False,
#                                          valida=True,
#                                          observaciones='')
#            his.save()
#            his.actualizar()
#            print linea
#       else:
#           print "FALLO"
#           data = [[row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8],row[9],row[10],row[11],row[12],row[13],row[14],row[15],row[16],row[17],row[18]]]
#           dataWriter.writerows(data)
#       linea+=1
#   else:
#       print "FALLO"
#       data = [[row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8],row[9],row[10],row[11],row[12],row[13],row[14],row[15],row[16],row[17],row[18]]]
#       dataWriter.writerows(data)

# fecha = date(2014, 11, 21)
# materias = MateriaAsignada.objects.filter(materia__nivel__periodo_id=5)
# for m in materias:
#     if not RubroDerecho.objects.filter(materiaasignada=m).exists():
#         rubro = Rubro(fecha=datetime.now().date(),
#                       iva=False,
#                       valor=2,
#                       inscripcion=m.matricula.inscripcion,
#                       cancelado=False,
#                       fechavence=fecha)
#         rubro.save()
#         rubroderecho = RubroDerecho(rubro=rubro,
#                                     materiaasignada=m)
#         rubroderecho.save()


# matriculas = Matricula.objects.filter(nivel__periodo__id=4)
# for matricula in matriculas:
#     for materia in matricula.materiaasignada_set.all():
#         rubro = Rubro(fecha=datetime.now().date(),
#                       iva=False,
#                       valor=2,
#                       inscripcion=matricula.inscripcion,
#                       cancelado=False,
#                       fechavence=materia.materia.fin)
#         rubro.save()
#         rubro = RubroOtro(rubro=rubro,
#                           tipo_id=TIPO_OTRO_RUBRO,
#                           descripcion='DERECHO EXAMEN: ' + unicode(materia.materia.asignatura))
#         rubro.save()


# 2 REVICION
# 1 INGRESO


# BUSQUEDA DE INSCRIPCIONES POR NOMBRE
# linea = 1
# nombre = ''
# for row in dataReader:
#     print linea
#     if Persona.objects.extra(where=["CONCAT(apellido1, ' ', apellido2, ' ', nombres) LIKE %s"], params=[row[0]]).exists():
#         persona = Persona.objects.extra(where=["CONCAT(apellido1, ' ', apellido2, ' ', nombres) LIKE %s"], params=[row[0]])[0]
#         inscripcion = persona.inscripcion()
#         if inscripcion:
#             ids = inscripcion.id
#         else:
#             ids = ''
#     else:
#         ids = ''
#         # if nombre != row[0]:
#         #     nombre = row[0]
#         #     data = [[row[0]]]
#         #     dataWriter.writerows(data)
#
#     data = [[row[0],ids,row[2],row[3],row[4]]]
#     dataWriter.writerows(data)
#     linea += 1



# INSCRIPCIONES SIN GRUPO
# linea = 1
# grupo = Group.objects.get(pk=ALUMNOS_GROUP_ID)
# for row in dataReader:
#     if linea > 1:
#         carrera = Carrera.objects.get(pk=int(row[5]))
#         if not Inscripcion.objects.filter(persona__cedula=row[10], carrera=carrera).exists():
#             persona = Persona(nombres = row[8],
#                               apellido1 = row[6],
#                               apellido2 = row[7],
#                               cedula = row[10],
#                               pasaporte = '',
#                               nacimiento = convertirfecha(row[11]),
#                               sexo_id = int(row[9]),
#                               telefono = row[18],
#                               telefono_conv = row[17],
#                               direccion = row[16],
#                               pais_id=int(row[14]),
#                               email = row[27],
#                               emailinst = '')
#             persona.save()
#             persona.cambiar_clave()
#             username = calculate_username(persona)
#             password = DEFAULT_PASSWORD
#             user = User.objects.create_user(username, persona.email, password)
#             user.save()
#             grupo.user_set.add(user)
#             grupo.save()
#             persona.usuario = user
#             # persona.emailinst = user.username+'@'+EMAIL_DOMAIN
#             persona.save()
#             if Especialidad.objects.filter(nombre=row[25]).exists():
#                 espepecialidad = Especialidad.objects.filter(nombre=row[25])[0]
#             else:
#                 espepecialidad = Especialidad(nombre=row[25])
#                 espepecialidad.save()
#             try:
#                 fechains = convertirfecha(row[26])
#             except Exception as ex:
#                 fechains = datetime.now().date()
#             inscripcion = Inscripcion(persona = persona,
#                                       fecha = fechains,
#                                       carrera = carrera,
#                                       modalidad_id = int(row[1]),
#                                       sesion_id = int(row[3]),
#                                       sede_id = 1,
#                                       colegio = row[24],
#                                       especialidad_id = 251,
#                                       identificador = '',
#                                       tienediscapacidad = False,
#                                       cumplimiento = False)
#             inscripcion.save()
#             if inscripcion.inscripcionmalla_set.exists():
#                 imv = inscripcion.inscripcionmalla_set.all()
#                 for im in imv:
#                     im.delete()
#             inscripcion.malla_inscripcion()
#             print linea
#         else:
#             data = [[row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8],row[9],row[10],row[11],row[12],row[13],row[14],row[15],row[16],row[17],row[18],row[19],row[20],row[21],row[22],row[23],row[24],row[25],row[26],row[27]]]
#             # dataWriter.writerows(data)
#             print linea
#     linea += 1






#BUSQUEDA DE INSCRIPCIONES POR NOMBRE
# linea = 1
# gru = Group.objects.get(pk=2)
# for row in dataReader:
#     if not Inscripcion.objects.filter(persona__cedula=row[0]).exists():
#         persona = Persona(nombres = row[3],
#                           apellido1 = row[1],
#                           apellido2 = row[2],
#                           cedula = row[0],
#                           pasaporte = '',
#                           nacimiento = convertirfecha(row[7]),
#                           sexo_id = 1 if row[11] == "F" else 2,
#                           telefono = row[5],
#                           telefono_conv = row[6],
#                           direccion = row[4],
#                           nacionalidad=row[10],
#                           email = '',
#                           emailinst = '')
#         persona.save()
#         persona.cambiar_clave()
#         username = calculate_username(persona)
#         password = DEFAULT_PASSWORD
#         user = User.objects.create_user(username, persona.email, password)
#         user.save()
#         gru.user_set.add(user)
#         gru.save()
#         persona.usuario = user
#         # persona.emailinst = user.username+'@'+EMAIL_DOMAIN
#         persona.save()
#         if Grupo.objects.filter(id=int(row[13])).exists():
#             grupo = Grupo.objects.get(pk=int(row[13]))
#         else:
#             raise
#         # if Especialidad.objects.filter(nombre=row[15]).exists():
#         #     espepecialidad = Especialidad.objects.filter(nombre=row[15])[0]
#         # else:
#         #     espepecialidad = Especialidad(nombre=row[15])
#         #     espepecialidad.save()
#         try:
#             fechains = convertirfecha(row[14])
#         except Exception as ex:
#             fechains = datetime.now().date()
#         inscripcion = Inscripcion(persona = persona,
#                                   fecha = fechains,
#                                   carrera = grupo.carrera,
#                                   modalidad = grupo.modalidad,
#                                   sesion = grupo.sesion,
#                                   sede = grupo.sede,
#                                   colegio = '',
#                                   especialidad_id = 183,
#                                   identificador = '',
#                                   tienediscapacidad = False,
#                                   cumplimiento = False)
#         inscripcion.save()
#         if inscripcion.inscripcionmalla_set.exists():
#             imv = inscripcion.inscripcionmalla_set.all()
#             for im in imv:
#                 im.delete()
#         inscripcion.malla_inscripcion()
#
#         if inscripcion.inscripciongrupo_set.exists():
#             ing = inscripcion.inscripciongrupo_set.all()
#             for ig in ing:
#                 ig.delete()
#         ig = InscripcionGrupo(inscripcion=inscripcion,
#                               grupo=grupo,
#                               activo=True)
#         ig.save()
#         print linea
#     else:
#         data = [[row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8],row[9],row[10],row[11],row[12],row[13],row[14]]]
#         dataWriter.writerows(data)
#         print linea
#     linea += 1
#
#  SOLO NOMBRES SIN DATOS
#
# linea = 1
# for row in dataReader:
#     print linea
#     if Persona.objects.filter(nombres=row[3].strip(), apellido1=row[1].strip(), apellido2=row[2].strip()).exists():
#         print "Existe"
#     else:
#         persona = Persona(nombres = remover_caracteres_especiales(to_unicode(row[3])),
#                               apellido1 = remover_caracteres_especiales(to_unicode(row[1])),
#                               apellido2 = remover_caracteres_especiales(to_unicode(row[2])),
#                               cedula = '0000000000',
#                               pasaporte = '',
#                               nacimiento = date(1980,1,1),
#                               sexo_id = 1,
#                               telefono = '',
#                               telefono_conv = '',
#                               email = '',
#                               emailinst = '')
#         persona.save()
#         persona.cambiar_clave()
#         username = calculate_username(persona)
#         password = DEFAULT_PASSWORD
#         user = User.objects.create_user(username, persona.email, password)
#         user.save()
#         persona.usuario = user
#         persona.emailinst = user.username+'@'+EMAIL_DOMAIN
#         persona.save()
#         grupo = Grupo.objects.get(pk=74)
#         espepecialidad = Especialidad.objects.get(pk=183)
#         fechains = datetime.now().date()
#         inscripcion = Inscripcion(persona = persona,
#                                   fecha = fechains,
#                                   carrera = grupo.carrera,
#                                   modalidad = grupo.modalidad,
#                                   sesion = grupo.sesion,
#                                   sede = grupo.sede,
#                                   colegio = '',
#                                   especialidad = espepecialidad,
#                                   identificador = '',
#                                   tienediscapacidad = False,
#                                   cumplimiento = False)
#         inscripcion.save()
#         if inscripcion.inscripcionmalla_set.exists():
#             imv = inscripcion.inscripcionmalla_set.all()
#             for im in imv:
#                 im.delete()
#         inscripcion.malla_inscripcion()
#
#         if inscripcion.inscripciongrupo_set.exists():
#             ing = inscripcion.inscripciongrupo_set.all()
#             for ig in ing:
#                 ig.delete()
#         ig = InscripcionGrupo(inscripcion=inscripcion,
#                               grupo=grupo,
#                               activo=True)
#         ig.save()
#     linea += 1




# if tipopasada ==1:
#     linea = 1
#     for row in dataReader:
#         print linea
#         if Persona.objects.filter(cedula=row[3]).exists() or row[3]=='':
#             print "FALLO"
#             try:
#                 data = [[row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8],row[9],row[10],row[11],row[12],row[13]]]
#             except Exception as ex:
#                 data = [[row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8],row[9],row[10],row[11],row[12]]]
#             dataWriter.writerows(data)
#         else:
#             persona = Persona(nombres = remover_caracteres_especiales(to_unicode(row[2])),
#                               apellido1 = remover_caracteres_especiales(to_unicode(row[0])),
#                               apellido2 = remover_caracteres_especiales(to_unicode(row[1])),
#                               cedula = row[3],
#                               pasaporte = '',
#                               nacimiento = convertirfecha(row[4]),
#                               sexo_id = 1,
#                               telefono = row[7],
#                               telefono_conv = row[6],
#                               email = remover_caracteres_especiales(to_unicode(row[8])),
#                               emailinst = '')
#             persona.save()
#             persona.cambiar_clave()
#             username = calculate_username(persona)
#             password = DEFAULT_PASSWORD
#             user = User.objects.create_user(username, persona.email, password)
#             user.save()
#             persona.usuario = user
#             persona.emailinst = user.username+'@'+EMAIL_DOMAIN
#             persona.save()
#             grupo = Grupo.objects.get(pk=int(row[12]))
#             if Especialidad.objects.filter(nombre=remover_caracteres_especiales(to_unicode(row[10]))).exists():
#                 espepecialidad = Especialidad.objects.filter(nombre=remover_caracteres_especiales(to_unicode(row[10])))[0]
#             else:
#                 if remover_caracteres_especiales(to_unicode(row[10])).__len__() > 0:
#                     espepecialidad = Especialidad(nombre=remover_caracteres_especiales(to_unicode(row[10])))
#                     espepecialidad.save()
#                 else:
#                     espepecialidad = Especialidad.objects.get(pk=183)
#
#             try:
#                 fechains = convertirfecha(row[13])
#             except Exception as ex:
#                 fechains = datetime.now().date()
#
#             inscripcion = Inscripcion(persona = persona,
#                                       fecha = fechains,
#                                       carrera = grupo.carrera,
#                                       modalidad = grupo.modalidad,
#                                       sesion = grupo.sesion,
#                                       sede = grupo.sede,
#                                       colegio = remover_caracteres_especiales(to_unicode(row[9])),
#                                       especialidad = espepecialidad,
#                                       identificador = '',
#                                       tienediscapacidad = False,
#                                       cumplimiento = False)
#             inscripcion.save()
#             if inscripcion.inscripcionmalla_set.exists():
#                 imv = inscripcion.inscripcionmalla_set.all()
#                 for im in imv:
#                     im.delete()
#             inscripcion.malla_inscripcion()
#
#             if inscripcion.inscripciongrupo_set.exists():
#                 ing = inscripcion.inscripciongrupo_set.all()
#                 for ig in ing:
#                     ig.delete()
#             ig = InscripcionGrupo(inscripcion=inscripcion,
#                                   grupo=grupo,
#                                   activo=True)
#             ig.save()
#         linea += 1

# elif tipopasada == 2:
#     linea = 0
#     for row in dataReader:
#         if Inscripcion.objects.filter(cedula=row[3]).exists():
#             print "FALLO"
#             data = [[row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8],row[9],row[10],row[11],row[12],row[13]]]
#             dataWriter.writerows(data)
#




#
# inscripcion = Inscripcion.objects.filter(identificador=row[0])
# if inscripcion:
#     asignatura = Asignatura.objects.filter(nombre=row[3])
#     if asignatura:
#         fecha = convertirfecha(row[6])
#         if not HistoricoRecordAcademico.objects.filter(inscripcion=inscripcion[0], asignatura=asignatura[0], fecha=fecha).exists():
#             historico = HistoricoRecordAcademico(inscripcion=inscripcion[0],
#                                                  asignatura=asignatura[0],
#                                                  nota=0,
#                                                  asistencia=0,
#                                                  fecha=fecha,
#                                                  aprobada=False,
#                                                  convalidacion=False,
#                                                  pendiente=False,
#                                                  creditos=0,
#                                                  homologada=False,
#                                                  valida=False,
#                                                  observaciones=row[1])
#             historico.save()
#             historico.actualizar()


#
# docentes = Profesor.objects.all()
# for profesor in docentes:
#     if profesor.profesormateria_set.exists():
#         pm = profesor.profesormateria_set.all().order_by('-id')[0]
#         coordinacion = pm.materia.nivel.nivellibrecoordinacion_set.all()[0].coordinacion
#         profesor.coordinacion = coordinacion
#         profesor.save()
#         print unicode(coordinacion) + " - " + unicode(profesor)
#     else:
#         print "SIN COORDINACION - " + unicode(profesor)



# records = HistoricoRecordAcademico.objects.filter(aprobada=True, valida=True, convalidacion=False, homologada=False)
# for record in records:
#     print record.id
#     if not record.existe_en_malla():
#         record.valida = False
#         record.save()
#
# records = RecordAcademico.objects.filter(aprobada=True, valida=True, convalidacion=False, homologada=False)
# for record in records:
#     print record.id
#     if not record.existe_en_malla():
#         record.valida = False
#         record.save()



#
# records = RecordAcademico.objects.filter(homologada=True)
# for r in records:
#     try:
#         homologacion = HomologacionInscripcion.objects.filter(record=r)[0]
#         try:
#             historico = r.historicorecordacademico_set.filter(fecha=r.fecha, nota=r.nota)[0]
#             homologacion.historico = historico
#             homologacion.save()
#         except Exception as ex:
#             pass
#             r.actualizar()
#             print "no historico: "+str(r.id)
#     except Exception as ex:
#         pass
#         print "no tiene: "+str(r.id)
#
#

#hist = HistoricoRecordAcademico.objects.filter(aprobada=False)
#for h in hist:
#    h.save()

#record = RecordAcademico.objects.all()
#for r in record:
#    historico = HistoricoRecordAcademico.objects.filter(inscripcion=r.inscripcion, asignatura=r.asignatura)
#    print r
#    if historico.count() > 0:
#
#        for h in historico:
#            h.recordacademico = r
#            h.save()
#    else:
#        print "ARREGLO xxxxxxxxxxxxxxxxxxxxxxxxx:"+unicode(r)
#        r.actualizar()

#histo = HistoricoRecordAcademico.objects.filter(recordacademico=None)
#for h in histo:
#    print h.id
#    try:
#        record = RecordAcademico.objects.filter(inscripcion=h.inscripcion, asignatura=h.asignatura)[:1].get()
#        h.recordacademico = record
#        h.save()
#    except Exception as ex:
#        print "ACTUALIZADA......"
#        h.actualizar()
#
# graduados = Graduado.objects.filter()
# for g in graduados:
#     if g.inscripcion.creditos() < g.inscripcion.malla_inscripcion().malla.creditos_completar:
#         print unicode(g.inscripcion) + " - " + g.inscripcion.creditos().__str__() + " " + unicode(g.inscripcion.malla_inscripcion().malla)


#for ins in Inscripcion.objects.all():
#    ins.persona.nombres = ins.persona.nombres.replace(u"Ñ", "N").replace(u"Á", "A").replace(u"É", "E").replace(u"Í", "I").replace(u"Ó", "O").replace(u"Ú", "U").lstrip()
#    ins.persona.apellido1 = ins.persona.apellido1.replace(u"Ñ", "N").replace(u"Á", "A").replace(u"É", "E").replace(u"Í", "I").replace(u"Ó", "O").replace(u"Ú", "U").lstrip()
#    ins.persona.apellido2 = ins.persona.apellido2.replace(u"Ñ", "N").replace(u"Á", "A").replace(u"É", "E").replace(u"Í", "I").replace(u"Ó", "O").replace(u"Ú", "U").lstrip()
#    ins.save()
#    print unicode(ins)+" - "+ str(ins.id)
#    unicodedata.normalize('NFKD', ins.persona.apellido2).encode('ascii', 'ignore')

#dataReader = csv.reader(open(csv_filepathname, "rU"), delimiter=',')
#ERRORES = 0

#
# linea = 1
# for row in dataReader:
#     crear = False
#     if Persona.objects.filter(cedula=row[1]).exists():
#         print "EXISTENTE --> " + row[1]
#     else:
#         persona = Persona(nombres=row[2].strip(),
#                           apellido1=row[3].strip(),
#                           apellido2=row[4].strip(),
#                           cedula=row[1],
#                           pasaporte='',
#                           nacimiento=convertirfecha(row[5]),
#                           pais_id=int(row[6]),
#                           provincianac=row[7].strip(),
#                           cantonnac=row[8].strip(),
#                           parroquianac=row[9].strip(),
#                           sexo_id=int(row[10]),
#                           direccion=row[14].strip(),
#                           direccion2=row[15].strip(),
#                           num_direccion=row[16].strip(),
#                           sector=row[17].strip(),
#                           ciudad='',
#                           telefono=row[19].strip(),
#                           telefono_conv=row[20].strip(),
#                           email=row[18].strip(),
#                           emailinst='')
#         persona.save()
#         persona.cambiar_clave()
#         username = calculate_username(persona, estudiante=True)
#         generar_usuario(persona, username, ALUMNOS_GROUP_ID)
#         inscripcion = Inscripcion(persona=persona,
#                                  fecha=date(2015, 2, 1),
#                                  colegio='',
#                                  carrera_id=int(row[0]),
#                                  sede=Sede.objects.all()[0],
#                                  modalidad=Modalidad.objects.all()[0],
#                                  coordinacion=Coordinacion.objects.all()[0],
#                                  sesion=Sesion.objects.all()[0])
#         inscripcion.save()
#         documentos = DocumentosDeInscripcion(inscripcion=inscripcion,
#                                             titulo=True,
#                                             acta=True,
#                                             cedula=True,
#                                             votacion=True,
#                                             actaconv=False,
#                                             partida_nac=False,
#                                             pre=False,
#                                             observaciones_pre='',
#                                             fotos=False)
#         documentos.save()
#         inscripciontipoinscripcion = InscripcionTipoInscripcion(inscripcion=inscripcion,
#                                                                 tipoinscripcion_id=TIPO_INSCRIPCION_INICIAL)
#         inscripciontipoinscripcion.save()
#         persona.crear_perfil(inscripcion=inscripcion)
#         inscripcion.persona.mi_perfil()
#         inscripcion.malla_inscripcion()
#         print persona.cedula
#         linea += 1

# dataReader = csv.reader(open(csv_filepathname, "rU"), delimiter=';')
# linea = 0
# for row in dataReader:
#     print linea
#     if not Asignatura.objects.filter(nombre=row[0]).exists():
#         asignatura = Asignatura(nombre=row[0])
#         asignatura.save()
#     linea += 1






#
#    if Asignatura.objects.filter(codigo=row[0])>1
#    inscrip = Inscripcion.objects.filter(persona__cedula=row[0])[:1].get()
#    data = [[inscrip.id.__str__(),row[0],row[1]]]
#    dataWriter.writerows(data)
#else:
#    data = [['ERROR',row[0],row[1]]]
#    dataWriter.writerows(data)
#    print 'ERROR CON NUMERO DE CEDULA DE '+row[1]
#    ERRORES +=1
#
# if ERRORES==0:
#     print 'PROCESO TERMINADO SIN INCONVENIENTES'
# else:
#     print 'PROCESO TERMINADO CON '+ERRORES.__str__() +' INCONVENIENTES'

# def convertir_fecha_corta(f):
#     s = f.split("/")
#     if int(s[2])<=20:
#         return datetime(2000+int(s[2]), int(s[0]), int(s[1]))
#     else:
#         return datetime(1900+int(s[2]), int(s[0]), int(s[1]))

# noexiste = 0
#failed = open('failed.txt', 'w')
#

# def fecha_simple(s):
#     return datetime(year=1900+int(s[s.rindex("/")+1:]), month=int(s[:s.index("/")]), day=int(s[s.index("/")+1:s.rindex("/")]))
# first = True

# matriculas = Matricula.objects.filter(nivel__cerrado=False)
#
# inscripcion = Inscripcion.objects.all()
#
# for ins in inscripcion:
#     print ins
#     historico = ins.historicorecordacademico_set.all()
#     # historico = ins.recordacademico_set.all()
#     malla = ins.malla_inscripcion().malla
#     for his in historico:
#         if his.esta_suspensa():
#             cre=0
#         elif  malla.asignaturamalla_set.filter(asignatura=his.asignatura).exists():
#             cre = malla.asignaturamalla_set.filter(asignatura=his.asignatura)[:1].get().creditos
#         else:
#             cre = 0
#         his.creditos = cre
#         his.save()
#         his.actualizar()
#
#     #for his in historico:
#     #    if ins.recordacademico_set.filter(asignatura=his.asignatura).count()>1:
#     #        print ins
#     #        print his.asignatura
#
#

#calculo
#as1 = Asignatura.objects.get(pk=182)
#estadistica
#as2 = Asignatura.objects.get(pk=44)

# fecha = date(2004,5,2)

#per = Persona.objects.all()
#prefijo = "04"
#prefijo_ciudad = "2"
#for p in per:
#    print p
#    if p.telefono_conv[:2] == "09" or p.telefono_conv[:2] == "08":
#        p.telefono = p.telefono_conv
#        p.telefono_conv = ""
#    if p.telefono[:2] != "09" and p.telefono[:2] != "08":
#        p.telefono_conv = p.telefono
#        p.telefono = ""
#
#    p.telefono_conv = re.sub("\D", "", p.telefono_conv)
#    if p.telefono_conv.__len__() > 9:
#        p.telefono_conv = p.telefono_conv[:9]
#    if p.telefono_conv.__len__() == 8:
#        p.telefono_conv = prefijo[0] + p.telefono_conv
#    if p.telefono_conv.__len__() == 7:
#        p.telefono_conv = prefijo + p.telefono_conv
#    if p.telefono_conv.__len__() == 6:
#        p.telefono_conv = prefijo + prefijo_ciudad + p.telefono_conv
#    if p.telefono_conv.__len__() < 6:
#        p.telefono_conv = ''
#    print p.telefono_conv
#
#    p.telefono = re.sub("\D", "", p.telefono)
#    if p.telefono.__len__() > 10:
#        p.telefono = p.telefono[:10]
#    if p.telefono.__len__() == 9:
#        p.telefono = p.telefono[0] + "9" + p.telefono[1:]
#    if p.telefono.__len__() < 9:
#        p.telefono = ''
#    print p.telefono
#
#    p.cedula = re.sub("\D", "", p.cedula)
#    p.pasaporte = re.sub("\D", "", p.pasaporte)
#    if p.cedula == p.pasaporte:
#        p.pasaporte = ""
#    if p.cedula.__len__() != 10:
#        p.pasaporte = p.cedula
#        p.cedula = ""
#    if p.pasaporte.__len__() == 10 and p.cedula=="":
#        p.cedula = p.pasaporte
#        p.pasaporte = ""
#
#    if not p.nacimiento:
#        p.nacimiento = date(1980, 1, 1)
#
#    if p.emailinst == "ycarrenop@gmail.com":
#        pass
#
#    if p.email and p.email.find(EMAIL_DOMAIN) > 0:
#        p.emailinst = p.email
#        p.email = ""
#
#    if p.emailinst and p.emailinst.find(EMAIL_DOMAIN) < 1:
#        p.emailinst = ""
#
#    if not p.nacionalidad:
#        if p.sexo_id == SEXO_FEMENINO:
#            p.nacionalidad = "ECUATORIANA"
#        else:
#            p.nacionalidad = "ECUATORIANO"
#            "ECUADOR"
#    if p.nacionalidad == "ECUADOR" or p.nacionalidad == "EC":
#        if p.sexo_id == SEXO_FEMENINO:
#            p.nacionalidad = "ECUATORIANA"
#        else:
#            p.nacionalidad = "ECUATORIANO"
#    if p.nacionalidad == "ECUATORIANO" and p.sexo_id == SEXO_FEMENINO:
#        p.nacionalidad = "ECUATORIANA"
#    if p.nacionalidad == "ECUATORIANA" and p.sexo_id == SEXO_MASCULINO:
#        p.nacionalidad = "ECUATORIANO"
#    if p.nacionalidad == "CUBANO" and p.sexo_id == SEXO_FEMENINO:
#        p.nacionalidad = "CUBANA"
#    if p.nacionalidad == "CUBANA" and p.sexo_id == SEXO_MASCULINO:
#        p.nacionalidad = "CUBANO"
#    if p.nacionalidad == "PERUANO" and p.sexo_id == SEXO_FEMENINO:
#        p.nacionalidad = "PERUANA"
#    if p.nacionalidad == "PERUANA" and p.sexo_id == SEXO_MASCULINO:
#        p.nacionalidad = "PERUANO"
#    if p.nacionalidad == "COLOMBIANO" and p.sexo_id == SEXO_FEMENINO:
#        p.nacionalidad = "COLOMBIANA"
#    if p.nacionalidad == "COLOMBIANA" and p.sexo_id == SEXO_MASCULINO:
#        p.nacionalidad = "COLOMBIANO"
#
#
#
#    p.save()



#for r in ins:
#    if r.persona.telefono_conv.__len__() != 9:


#print r.inscripcion
#print r.inscripcion.carrera
# masantigua = r.inscripcion.recordacademico_set.filter(asignatura=as1).order_by('fecha')[0]
# masantigua.asignatura = as2
# masantigua.save()
# hmasantigua = r.inscripcion.historicorecordacademico_set.filter(asignatura=as1).order_by('fecha')[0]
# hmasantigua.asignatura = as2
# hmasantigua.save()


# elif r.inscripcion.recordacademico_set.filter(asignatura=as1).count()==1 and r.inscripcion.recordacademico_set.filter(asignatura=as2).count()==0:
#     print r.inscripcion
#     pass
# else:
#     r.asignatura = as2
#     r.save()
#     his = r.inscripcion.historicorecordacademico_set.filter(asignatura=as1)
#     for h in his:
#         h.asignatura = as2
#         h.save()


# fecha = date(1990,12,1)
# reco = Inscripcion.objects.filter(recordacademico__fecha__lt=fecha).distinct()
# for r in reco:
#     print r

###
###   VERIFICAR RUBROS
###
# periodo = Periodo.objects.get(pk=78)
# matriculados = Matricula.objects.filter(nivel__id=244)
# hoy = datetime.now().date()
# for mat in matriculados:
#     rubro = Rubro(fecha = datetime.now().date(),
#                   iva = False,
#                   valor = 130,
#                   inscripcion = mat.inscripcion,
#                   cancelado = False,
#                   fechavence= date(2014,6,11))
#     rubro.save()
#     rubromatricula = RubroMatricula(rubro=rubro, matricula=mat)
#     rubromatricula.save()

# costomateria = mat.nivel.pagonivel_set.all()[0].valor
# if mat.becado:
#     pp = (100 - mat.porcientobeca) / 100.0
#     costomateria *= pp
#     costomateria = round(costomateria, 2)
#
# total_real = 0
# for cuota in mat.rubrocuota_set.all():
#     total_real += cuota.rubro.valor
#
# cantida_materias = 0
# for materia in mat.materiaasignada_set.all():
#     if not materia.esplan12():
#         cantida_materias += 1
# total_calculado = cantida_materias*costomateria
#
# if total_calculado!=total_real:
#     print str(mat.id) + " " +  mat.inscripcion.persona.nombre_completo_inverso() + " " + str(total_calculado) + " " + str (total_real)
#

# INCLUYE LAS MATERIAS DEL PRE A TODAS LAS MALLAS EXCEPTO A LAS QUE SE INDICA NO INCLUIR

# no_incluir=(36,35,37,38,41)
# materiaspre=AsignaturaMalla.objects.filter(malla__carrera__id=36)
# mallas = Malla.objects.all().exclude(carrera_id__in=(no_incluir))
# for m in mallas:
#     for materiapre in materiaspre:
#         a=AsignaturaMalla(malla=m,
#         asignatura=materiapre.asignatura,
#         nivelmalla=materiapre.nivelmalla,
#         ejeformativo=materiapre.ejeformativo,
#         horas=materiapre.horas,
#         creditos=materiapre.creditos,
#         rectora=materiapre.rectora,
#         identificacion=materiapre.identificacion)
#         a.save()

# ACTUALIZA EL ESTADO DE HISTORICO Y RECORD DE LOS PERIODOS INDICADOS

# periodos = (1)
# materias = MateriaAsignada.objects.filter(matricula__nivel__periodo__id=1)
# for m in materias:
#     if HistoricoRecordAcademico.objects.filter(inscripcion=m.matricula.inscripcion,asignatura=m.materia.asignatura,nota__gte=70,aprobada=False).exists():
#         historico=HistoricoRecordAcademico.objects.filter(inscripcion=m.matricula.inscripcion,asignatura=m.materia.asignatura,nota__gte=70,aprobada=False)
#         for h in historico:
#             print h
#             h.aprobada=True
#             if h.existe_en_malla():
#                 h.valida=True
#             h.creditos=m.materia.creditos
#             h.save()
#             h.actualizar()

# PERIODO 77 ACTUALIZAR SOLO LOS HISTORICOS DE LAS QUE CUMPLAN CON EL PORCENTAJE DE ASISTENCIA

# materias = MateriaAsignada.objects.filter(matricula__nivel__periodo_id=77)
# for m in materias:
#     if HistoricoRecordAcademico.objects.filter(inscripcion=m.matricula.inscripcion,asignatura=m.materia.asignatura,nota__gte=70,asistencia__gte=75,aprobada=False).exists():
#         historico=HistoricoRecordAcademico.objects.filter(inscripcion=m.matricula.inscripcion,asignatura=m.materia.asignatura,nota__gte=70,asistencia__gte=75,aprobada=False)
#         for h in historico:
#             print h
#             h.aprobada=True
#             if h.existe_en_malla():
#                 h.valida=True
#             h.creditos=m.materia.creditos
#             h.save()
#             h.actualizar()

# documentos = Documento.objects.filter(codigo__icontains='U-')
# for documento in documentos:
#     print documento.codigo
#     documento.codigo = documento.codigo[2:]
#     documento.save()
#     print documento.codigo

# # Actualizar los estados de los porcentajes de asistencia
# ma = MateriaAsignada.objects.filter(materia_id__gte=3564)
# for m in ma:
#     # m.asistenciafinal = m.porciento_asistencia()
#     m.save(actualiza=True)
#     m.actualiza_estado()
#     print 'procesado'
# print 'listo'

# ingles casos de noaplica modulos diferentes de 0
# h=HistoricoRecordAcademico.objects.filter(asignatura__modulo=True,noaplica=True).exclude(asignatura_id=782).exclude(observaciones__contains='SUFI').exclude(observaciones__contains='CERT').exclude(observaciones__contains='NO APLICA')
# for h1 in h:
#     if h1.inscripcion.inscripcionmalla_set.exists():
#         if h1.recordacademico.noaplica:
#             if not Egresado.objects.filter(inscripcion=h1.inscripcion_id).exists():
#                 print 'cedula:%s, carrera:%s ,modulo:%s, %s' % (h1.inscripcion.persona.cedula, h1.inscripcion.carrera.nombre ,h1.asignatura.nombre, h1.observaciones)
#                 h1.noaplica=False
#                 h1.aprobada=False
#                 h1.save()
#                 h1.actualizar()


# r = RecordAcademico.objects.filter(asignatura__modulo=True, aprobada=False).exclude(asignatura_id=782).distinct().order_by()
# for r1 in r:
#     id1=r1.inscripcion.id
#     asi1=r1.asignatura.id
#     if RecordAcademico.objects.filter(asignatura__modulo=True, aprobada=False, inscripcion_id=id1).exclude(asignatura_id=782).exclude(asignatura_id=asi1).exists():
#         print 'cedula:%s, carrera:%s ,modulo:%s' % (r1.inscripcion.persona.cedula, r1.inscripcion.carrera.nombre ,r1.asignatura.nombre )
# bandera=0
#
# #desmatricular un grupo de alumnos de una materia
# materiaasignada = MateriaAsignada.objects.filter(materia__in=[3720])
# request = Persona.objects.filter(pk=26995)[0]
# for ma in materiaasignada:
#     matricula = ma.matricula
#     matricula.eliminar_materia(ma, request)
#     matricula.actualizar_horas_creditos()
#     print matricula
# print 'listo'

# # contratos archivos
# n11=1
# contratopersona1 = ContratoPersona.objects.filter(contrato__anio=2016)
# for contratopersona in contratopersona1:
#     print n11
#     if n11 == 79:
#         pass
#     n11 += 1
#     nombre_plantilla = contratopersona.contrato.descripcion + ".docx"
#     nombre_contrato = str(contratopersona.id) + ".docx"
#
#     anio_plantilla = contratopersona.contrato.anio
#     direccion_plantilla = os.path.join(SITE_ROOT, 'media', 'contratos', 'plantillas', str(anio_plantilla))
#     direccion_contrato = os.path.join(SITE_ROOT, 'media', 'contratos', 'contrato')
#
#     filename_plantilla = os.path.join(direccion_plantilla, nombre_plantilla)
#     filename_contrato = os.path.join(direccion_contrato, nombre_contrato)
#
#     contratopersonadetalle1 = ContratoPersonaDetalle.objects.filter(contratopersona=contratopersona, status=True)
#     document = Document(filename_plantilla)
#     parrafo = document.paragraphs
#     cantidad_parrafo = parrafo.__len__()
#     for contratopersonadetalle in contratopersonadetalle1:
#         n = 0
#         campo = CamposContratos.objects.filter(pk=contratopersonadetalle.campos.id)[0]
#         if campo.script[:11] == 'JAVASCRIPT:':
#             campo1 = contratopersonadetalle.valor
#             if campo1 != '':
#                 valor = eval(campo.script[11:])
#                 campo_buscar = '{' + campo.descripcion.lower() + '|0}'
#                 campo_buscar1 = '{' + campo.descripcion.lower() + '|1}'
#                 for n in range(cantidad_parrafo):
#                     parrafo[n].text = parrafo[n].text.replace(campo_buscar, campo1)
#                     parrafo[n].text = parrafo[n].text.replace(campo_buscar1, valor)
#                     n += 1
#         else:
#             campo_buscar = '{' + campo.descripcion.lower() + '|0}'
#             for n in range(cantidad_parrafo):
#                 parrafo[n].text = parrafo[n].text.replace(campo_buscar, contratopersonadetalle.valor)
#                 n += 1
#     n = 0
#     persona_nombre = contratopersona.persona.nombre_titulo()
#     persona_cedula = contratopersona.persona.cedula
#     cadena1 = '{empleado|0}'
#     cadena2 = '{cedula|0}'
#     for n in range(cantidad_parrafo):
#         parrafo[n].text = parrafo[n].text.replace(cadena1, persona_nombre)
#         parrafo[n].text = parrafo[n].text.replace(cadena2, persona_cedula)
#         n += 1
#     document.save(filename_contrato)
#     contratopersona.archivo.name = "contratos/contrato/%s" % nombre_contrato
#     contratopersona.save(usuario=User.objects.get(pk=1))
#
#
#

# # graduados
# workbook = xlrd.open_workbook("graduados_eenfermeria.xlsx")
# sheet = workbook.sheet_by_index(0)
# n = 1
# for rowx in range(sheet.nrows):
#     print n
#     if n > 1:
#         row = sheet.row_values(rowx)
#         inscripcion = None
#         if Inscripcion.objects.filter(persona__cedula=row[0].strip(),carrera__id=int(row[1])).exists():
#             inscripcion = Inscripcion.objects.filter(persona__cedula=row[0].strip(),carrera__id=int(row[1]))[0]
#
#         if inscripcion:
#             if not Graduado.objects.filter(inscripcion=inscripcion,inscripcion__carrera__id=int(row[1])).exists():
#                 graduado = Graduado(inscripcion=inscripcion,
#                                     tematesis='',
#                                     notatesis=0,
#                                     notafinal=0,
#                                     fechagraduado=convertirfecha2(row[2].strip()),
#                                     registro='',
#                                     numeroactagrado='',
#                                     promediogrado=0,
#                                     sustentacion=0,
#                                     promediotitulacion=0,
#                                     pasantias=0,
#                                     vinculacion=0,
#                                     practicas=0,
#                                     notagraduacion=0)
#                 graduado.save()
#                 graduado.promediotitulacion = round((graduado.promediogrado + graduado.notatesis + graduado.sustentacion) / 3, 2)
#                 if graduado.vinculacion == 0 and graduado.practicas == 0:
#                     graduado.notagraduacion = round((graduado.notafinal + graduado.promediotitulacion + graduado.pasantias) / 3, 2)
#                 elif graduado.vinculacion != 0 and graduado.practicas != 0:
#                     graduado.notagraduacion = round((graduado.notafinal + graduado.promediotitulacion + graduado.pasantias + graduado.vinculacion + graduado.practicas) / 5,2)
#                 elif graduado.vinculacion != 0:
#                     graduado.notagraduacion = round(
#                         (graduado.notafinal + graduado.promediotitulacion + graduado.pasantias + graduado.vinculacion) / 4, 2)
#                 elif graduado.practicas != 0:
#                     graduado.notagraduacion = round((graduado.notafinal + graduado.promediotitulacion + graduado.pasantias + graduado.practicas) / 4, 2)
#                 graduado.save()
#             else:
#                 graduado = Graduado.objects.filter(inscripcion=inscripcion,inscripcion__carrera__id=int(row[1]))[0]
#                 graduado.fechagraduado=convertirfecha2(row[2].strip())
#                 graduado.save()
#                 print 'repetido : %s' % row[0].strip()
#         else:
#             print 'no esta : %s' % row[0].strip()
#     n += 1
#
# print 'listo'





# # crear profesor
# workbook = xlrd.open_workbook("crearprofesor.xlsx")
# sheet = workbook.sheet_by_index(0)
# n = 1
# for rowx in range(sheet.nrows):
#     if n > 1:
#         row = sheet.row_values(rowx)
#
#         if Persona.objects.filter(cedula=row[0].strip()).exists():
#             persona = Persona.objects.filter(cedula=row[0].strip())[0]
#             if not persona.profesor_set.exists():
#                 profesor = Profesor(persona=persona,
#                                     activo=True,
#                                     fechaingreso=datetime.now().date(),
#                                     coordinacion=Coordinacion.objects.all()[0],
#                                     dedicacion=TiempoDedicacionDocente.objects.all()[0])
#                 profesor.save()
#                 grupo = Group.objects.get(pk=PROFESORES_GROUP_ID)
#                 grupo.user_set.add(persona.usuario)
#                 grupo.save()
#                 persona.crear_perfil(profesor=profesor)
#                 print 'creo docente : %s' % (row[0].strip())
#             else:
#                 print 'Ya existe este docente : %s' % (row[0].strip())
#         else:
#             print 'no existe este numero de cedula en persona : %s' % (row[0].strip())
#     n += 1
# print 'listo'
#


# # # Practicas PreProfesionales
# workbook = xlrd.open_workbook("practicaspreprofesionales7.xlsx")
# sheet = workbook.sheet_by_index(0)
# n = 1
# for rowx in range(sheet.nrows):
#     if n > 1:
#         row = sheet.row_values(rowx)
#         inscripcion = None
#         if Inscripcion.objects.filter(Q(persona__cedula=row[0].strip()) | Q(persona__pasaporte=row[0].strip()), carrera__id=int(row[1])).exists():
#            inscripcion = Inscripcion.objects.filter(Q(persona__cedula=row[0].strip()) | Q(persona__pasaporte=row[0].strip()), carrera__id=int(row[1]))[0]
#         if inscripcion:
#             fechadesde = None
#             if row[6].strip() <> '':
#                 fechadesde = convertirfecha2(row[6].strip())
#             fechahasta = None
#             if row[7].strip() <> '':
#                 fechahasta = convertirfecha2(row[7].strip())
#             if not PracticasPreprofesionalesInscripcion.objects.filter(inscripcion=inscripcion, institucion=row[3].strip()).exists():
#                 practicaspreprofesionalesinscripcion = PracticasPreprofesionalesInscripcion(inscripcion=inscripcion,
#                                                                                             numerohora=int(row[2]),
#                                                                                             institucion=row[3].strip(),
#                                                                                             tipoinstitucion=int(row[4]),
#                                                                                             sectoreconomico=int(row[5]),
#                                                                                             fechadesde=fechadesde,
#                                                                                             fechahasta=fechahasta)
#                 practicaspreprofesionalesinscripcion.save()
#                 print 'listo cedula %s' % (row[0].strip())
#             else:
#                 practicaspreprofesionalesinscripcion = PracticasPreprofesionalesInscripcion.objects.filter(inscripcion=inscripcion, institucion=row[3].strip())[0]
#                 practicaspreprofesionalesinscripcion.fechadesde = fechadesde
#                 practicaspreprofesionalesinscripcion.fechahasta = fechahasta
#                 practicaspreprofesionalesinscripcion.save()
#                 print 'listo cedula modificado %s' % (row[0].strip())
#                 # print 'repetido cedula %s' % (row[0].strip())
#         else:
#             # inscripcion = None
#             # contar = 0
#             # if Inscripcion.objects.filter(Q(persona__cedula=row[0].strip()) | Q(persona__pasaporte=row[0].strip())).exists():
#             #     contar = Inscripcion.objects.filter(Q(persona__cedula=row[0].strip()) | Q(persona__pasaporte=row[0].strip())).count()
#             #     if contar == 1:
#             #         inscripcion = Inscripcion.objects.filter(Q(persona__cedula=row[0].strip()) | Q(persona__pasaporte=row[0].strip()))[0]
#             #         if not PracticasPreprofesionalesInscripcion.objects.filter(inscripcion=inscripcion).exists():
#             #             practicaspreprofesionalesinscripcion = PracticasPreprofesionalesInscripcion(inscripcion=inscripcion,
#             #                                                                                         institucion=row[2].strip())
#             #             practicaspreprofesionalesinscripcion.save()
#             #             print 'listo cedula %s' % (row[0].strip())
#             #         else:
#             #             print 'repetido cedula %s' % (row[0].strip())
#             #     else:
#             #         print 'mas de una carrera cedula %s' % (row[0].strip())
#             # else:
#             print 'no existe esta cedula %s' % (row[0].strip())
#     n += 1
# print 'listo'

# # Practicas VINCULACION
# workbook = xlrd.open_workbook("vinculacion.xls")
# sheet = workbook.sheet_by_index(0)
# n = 1
# for rowx in range(sheet.nrows):
#     if n > 1:
#         row = sheet.row_values(rowx)
#         persona = None
#         if Persona.objects.filter(Q(cedula=row[1].strip()) | Q(pasaporte=row[1].strip())).exists():
#             persona = Persona.objects.filter(Q(cedula=row[1].strip()) | Q(pasaporte=row[1].strip()))[0]
#         if persona:
#             if not VinculacionInscripcionAux.objects.filter(persona=persona).exists():
#                 vinculacioninscripcionaux = VinculacionInscripcionAux(persona=persona,
#                                                                       tipovinculacion=int(row[0]))
#                 vinculacioninscripcionaux.save()
#                 print 'listo cedula %s' % (row[1].strip())
#             else:
#                 print 'repetido cedula %s' % (row[1].strip())
#         else:
#             print 'no existe esta cedula %s' % (row[1].strip())
#     n += 1
# print 'listo'

# # Practicas persona proveedor
# workbook = xlrd.open_workbook("proveedor.xlsx")
# sheet = workbook.sheet_by_index(3)
# n = 1
# for rowx in range(sheet.nrows):
#     print n
#     if n > 1:
#         row = sheet.row_values(rowx)
#         persona = None
#         if not Persona.objects.filter(Q(cedula=row[0].strip()) | Q(pasaporte=row[0].strip()) | Q(ruc=row[0].strip())).exists():
#             provincia = None
#             if row[9].strip() != '':
#                 provincia = Provincia.objects.filter(pk=int(row[9].strip()))[0]
#
#             cedula = ''
#             ruc = ''
#             pasaporte=''
#             tipo_persona=1
#             if len(row[0].strip())>10:
#                 ruc=row[0].strip()
#                 tipo_persona=2
#             elif len(row[0].strip())==10:
#                 cedula=row[0].strip()
#             else:
#                 pasaporte=row[0].strip()
#
#             proveedor = Persona(cedula=cedula,
#                                 ruc=ruc,
#                                 pasaporte=pasaporte,
#                                 nombres=row[2].strip(),
#                                 direccion=row[3].strip(),
#                                 telefono=row[7].strip(),
#                                 telefono_conv=row[6].strip(),
#                                 email=row[5].strip(),
#                                 pais_id=int(row[8].strip()),
#                                 provincia=provincia,
#                                 nacimiento=datetime.now().date(),
#                                 tipopersona=tipo_persona)
#             proveedor.save()
#             externo = Externo(persona=proveedor,
#                               nombrecomercial=row[1].strip(),
#                               nombrecontacto=row[4].strip())
#             externo.save()
#         else:
#             print 'existe esta cedula %s' % (row[0].strip())
#     n += 1
# print 'listo'
#
# workbook = xlrd.open_workbook("proveedor.xlsx")
# sheet = workbook.sheet_by_index(1)
# n = 1
# for rowx in range(sheet.nrows):
#     print n
#     if n > 1:
#         row = sheet.row_values(rowx)
#         persona = None
#         if Persona.objects.filter(Q(cedula=row[0].strip()) | Q(pasaporte=row[0].strip()) | Q(ruc=row[0].strip())).exists():
#             proveedor=Persona.objects.filter(Q(cedula=row[0].strip()) | Q(pasaporte=row[0].strip()) | Q(ruc=row[0].strip()))[0]
#             if Externo.objects.filter(persona=proveedor).exists():
#                 externo=Externo.objects.filter(persona=proveedor)[0]
#                 externo.nombrecontacto=row[1].strip()
#                 externo.save()
#     n += 1
# print 'listo'

#
# for personas in Persona.objects.filter(perfilusuario__isnull=True).order_by('-id'):
#     print personas.id
#     if Externo.objects.filter(persona=personas).exists():
#         externos = Externo.objects.filter(persona=personas)[0]
#         personas.crear_perfil(externo=externos)
# print 'listo'



# for personas in Persona.objects.filter(perfilusuario__isnull=True).order_by('-id'):
#     print personas
#     if Profesor.objects.filter(persona=personas).exists():
#         profesor = Profesor.objects.filter(persona=personas)[0]
#         personas.crear_perfil(profesor=profesor)
# print 'listo'


#  # calcular_nivel
# matriculas = Matricula.objects.filter(nivel__periodo__id=14).order_by('-id')
# # cursor = connection.cursor()
# for matricula in matriculas:
#     print(matricula.id)
#     matricula.calcula_nivel()
#     matricula.actualiza_matricula()
# print('listo')


# # pago tesoreria
# rubros = Rubro.objects.filter(status=True, cancelado=False).order_by('-id')
# for rubro in rubros:
#     print rubro.id
#     rubro.save(usuario=rubro.usuario_creacion)


# # recaudaciones
# workbook = xlrd.open_workbook("deudaspostgrado.xlsx")
# sheet = workbook.sheet_by_index(0)
# n = 1
# codigo = ''
# for rowx in range(sheet.nrows):
#     print n
#     if n > 1:
#         row = sheet.row_values(rowx)
#         persona = None
#         if Persona.objects.filter(Q(cedula=row[2].strip()) | Q(pasaporte=row[2].strip()) | Q(ruc=row[2].strip())).exists():
#             persona = Persona.objects.filter(Q(cedula=row[2].strip()) | Q(pasaporte=row[2].strip()) | Q(ruc=row[2].strip()))[0]
#
#         usuario = None
#         if Persona.objects.filter(Q(cedula=row[12].strip()) | Q(pasaporte=row[12].strip()) | Q(ruc=row[12].strip())).exists():
#             usuario = Persona.objects.filter(Q(cedula=row[12].strip()) | Q(pasaporte=row[12].strip()) | Q(ruc=row[12].strip()))[0].usuario
#
#         tiporubro = None
#         if TipoOtroRubro.objects.filter(pk=int(row[1].strip())).exists():
#             tiporubro = TipoOtroRubro.objects.filter(pk=int(row[1].strip()))[0]
#
#         periodo = None
#         matricula = None
#         if row[14].strip() != '':
#             periodo = Periodo.objects.filter(pk=int(row[14].strip()))[0]
#             matricula = Matricula.objects.filter(nivel__periodo = periodo, inscripcion__persona__cedula=row[2].strip())[0]
#
#         cancelado=False
#         if row[11].strip()=='pagado':
#             cancelado=True
#
#         if codigo != row[0].strip():
#             codigo = row[0].strip()
#             rubroaux = None
#         else:
#             rubroaux = rubro
#
#         rubro = Rubro(tipo_id=int(row[1].strip()),
#                       persona = persona,
#                       relacionados = rubroaux,
#                       matricula = matricula,
#                       # contratorecaudacion = None,
#                       nombre = tiporubro.nombre,
#                       cuota = int(row[3]),
#                       fecha = convertirfecha2(row[4].strip()),
#                       fechavence = convertirfecha2(row[5].strip()),
#                       valor = float(row[6]),
#                       iva_id = 1,
#                       valoriva = 0,
#                       valortotal = float(row[6]),
#                       saldo = float(row[6]),
#                       cancelado = cancelado)
#         rubro.save(usuario=usuario)
#         Rubro.objects.filter(pk=rubro.id).update(fecha_creacion=convertirfechahora(row[13].strip()))
#
#         if cancelado:
#             pago = Pago(rubro=rubro,
#                         fecha=convertirfecha2(row[4].strip()),
#                         subtotal0=float(row[6]),
#                         subtotaliva=0,
#                         iva=0,
#                         valordescuento=0,
#                         valortotal=float(row[6]),
#                         efectivo=True)
#             pago.save(usuario=usuario)
#             Pago.objects.filter(pk=pago.id).update(fecha_creacion=convertirfechahora(row[13].strip()))
#
#     n += 1
# print 'listo'



# for personas in Persona.objects.filter(apellido1='', ruc='', pasaporte=''):
#     nombreimportacion = personas.nombres.strip()
#     print personas.nombres.strip()
#     if len(nombreimportacion.split(' ')) >= 3:
#         importacionnombres = ''
#         for x in nombreimportacion.split(' ')[:2]:
#             importacionnombres += x + ' '
#         importacionnombres = importacionnombres.strip()
#         importacionapellido1 = nombreimportacion.split(' ')[2]
#         importacionapellido2 = nombreimportacion.split(' ')[3]
#         personas.nombres=importacionnombres
#         personas.apellido1=importacionapellido1
#         personas.apellido2=importacionapellido2
#         personas.save()
#     else:
#         importacionnombres = nombreimportacion.split(' ')[1]
#         importacionapellido1 = nombreimportacion.split(' ')[0]
#         importacionapellido2 = ''


# # record
# workbook = xlrd.open_workbook("enfermeriarecord.xlsx")
# sheet = workbook.sheet_by_index(0)
# n = 1
# for rowx in range(sheet.nrows):
#     print n
#     if n == 90:
#         pass
#     if n > 1:
#         row = sheet.row_values(rowx)
#         promedio = 0
#         if Inscripcion.objects.filter(persona__cedula=row[2].strip(), carrera__id=int(row[4])).exists():
#             inscripcion = Inscripcion.objects.filter(persona__cedula=row[2].strip(), carrera__id=int(row[4]))[0]
#             promedio = inscripcion.promedio_record()
#
#         data = [[row[0], row[1], row[2], promedio]]
#         dataWriter.writerows(data)
#     n +=1
# print 'listo'


# # sacar alumnos que le falta materia de ingles por aprobar
# n = 0
# for inscripcion in Inscripcion.objects.filter(matricula__nivel__periodo__id=7):
#     n +=1
#     # sacar los modulo de la malla
#     modulos = inscripcion.carrera.malla().modulomalla_set.all().exclude(asignatura__id__in=[782,1818,1819])
#     asi = []
#     for m in modulos:
#         asi.append(m.asignatura_id)
#     record = RecordAcademico.objects.filter(inscripcion=inscripcion, asignatura__in=asi)
#     for modu in modulos:
#         if not record.filter(asignatura=modu.asignatura, aprobada=True).exists():
#             alumno = inscripcion.persona.nombre_completo_inverso().replace(u"Ñ", "N").replace(u"Á", "A").replace(u"É", "E").replace(u"Í", "I").replace(u"Ó", "O").replace(u"Ú", "U").lstrip()
#             carrera = inscripcion.carrera.nombre.replace(u"Ñ", "N").replace(u"Á", "A").replace(u"É", "E").replace(u"Í", "I").replace(u"Ó", "O").replace(u"Ú", "U").lstrip()
#             nivelmalla = inscripcion.matricula_set.all()[0].nivelmalla.nombre
#             nivelcredito = inscripcion.inscripcionnivel_set.all()[0].nivel.nombre
#             data = [[carrera,alumno, modu.asignatura, nivelmalla, nivelcredito]]
#             dataWriter.writerows(data)
#     print  inscripcion.persona.nombre_completo_inverso()
#     print n


# # Modulos de computacion - Notas Historicas
# workbook = xlrd.open_workbook("computacion7.xlsx")
# sheet = workbook.sheet_by_index(0)
# linea = 1
# for rowx in range(sheet.nrows):
#     if linea > 1:
#         row = sheet.row_values(rowx)
#         if Asignatura.objects.filter(nombre=row[3].upper().strip()).exists():
#             asignatura = Asignatura.objects.filter(nombre=row[3].upper().strip())[0]
#         else:
#             asignatura = Asignatura(nombre=row[3].upper().strip(),
#                                     modulo=True)
#             asignatura.save()
#         carrera = Carrera.objects.get(pk=int(row[1]))
#
#         if Inscripcion.objects.filter(persona__cedula=row[0].strip(), carrera=carrera).exists():
#             inscripcion = Inscripcion.objects.filter(persona__cedula=row[0].strip(), carrera=carrera)[0]
#             if inscripcion.historicorecordacademico_set.filter(asignatura=asignatura).exists():
#                 if inscripcion.historicorecordacademico_set.filter(asignatura=asignatura, fecha = convertirfecha2(row[6].strip())).exists():
#                     record = inscripcion.historicorecordacademico_set.filter(asignatura=asignatura, fecha = convertirfecha2(row[6].strip())).order_by('-fecha')[0]
#                 else:
#                     record = inscripcion.historicorecordacademico_set.filter(asignatura=asignatura).order_by('-fecha')[0]
#                 record.nota = Decimal(row[4]).quantize(Decimal('.01'))
#                 record.fecha = convertirfecha2(row[6].strip())
#                 record.asistencia = 100 if row[5].strip() == "A" else 70
#                 record.aprobada = True if row[5].strip() == "A" else False
#                 record.pendiente = False
#                 record.creditos = 0
#                 record.horas = 0
#                 record.observaciones = row[2].strip() + " (M)"
#                 record.homologada = False
#                 record.valida = True if row[5].strip() == "A" else False
#                 record.noaplica = False
#                 record.save()
#                 print "Actualizado: " + str(linea) + " (" + inscripcion.persona.cedula + ")"
#             else:
#                 if not inscripcion.inscripcionmalla_set.filter(malla__carrera=carrera).exists():
#                     malla = carrera.malla()
#                     inscripcionmalla = InscripcionMalla(inscripcion=inscripcion,
#                                                         malla=malla)
#                     inscripcionmalla.save()
#
#                 record = RecordAcademico(inscripcion=inscripcion,
#                                          asignatura=asignatura,
#                                          nota=Decimal(row[4]).quantize(Decimal('.01')),
#                                          asistencia=70,
#                                          fecha=convertirfecha2(row[6].strip()),
#                                          aprobada=True if row[5].strip() == "A" else False,
#                                          pendiente=False,
#                                          creditos=0,
#                                          horas=0,
#                                          observaciones=row[2].strip() + " (M)",
#                                          homologada=False,
#                                          noaplica = False,
#                                          valida=True if row[5].strip() == "A" else False)
#                 record.save()
#                 print "Creado: " + str(linea) + " (" + inscripcion.persona.cedula + ")"
#             record.actualizar()
#
#         else:
#             data = [[row[0],row[1],row[2],row[3],row[4], row[5],"R"]]
#             dataWriter.writerows(data)
#             print "fallo"
#
#     linea += 1





# # Gastos personales
# workbook = xlrd.open_workbook("gastospersonales.xlsx")
# GastosPersonales.objects.all().delete()
# sheet = workbook.sheet_by_index(0)
# linea = 1
# for rowx in range(sheet.nrows):
#     if linea > 1:
#         print linea
#         row = sheet.row_values(rowx)
#         persona = None
#         if Persona.objects.filter(cedula=row[1].strip()).exists():
#             persona = Persona.objects.filter(cedula=row[1].strip())[0]
#
#         usuario = None
#         if Persona.objects.filter(cedula=row[30].strip()).exists():
#             usuario = Persona.objects.filter(cedula=row[30].strip())[0].usuario
#
#         periodo = None
#         if PeriodoGastosPersonales.objects.filter(anio=int(row[0].strip())).exists():
#             periodogastospersonales = PeriodoGastosPersonales.objects.filter(anio=int(row[0].strip()))[0]
#
#         if persona:
#             if not GastosPersonales.objects.filter(persona = persona, periodogastospersonales = periodogastospersonales, mes = int(row[29]), rmuproyectado = Decimal(row[5]).quantize(Decimal('.01')), rmupagado = Decimal(row[4]).quantize(Decimal('.01')), horasextraspagado = Decimal(row[6]).quantize(Decimal('.01')), horasextrasactual = Decimal(row[7]).quantize(Decimal('.01')), horasextrasproyectado = Decimal(row[8]).quantize(Decimal('.01')), otrosingresos = Decimal(row[9]).quantize(Decimal('.01')), totalingresos = Decimal(row[12]).quantize(Decimal('.01')), otrosgastos = Decimal(row[10]).quantize(Decimal('.01')), rebajasotros = Decimal(row[11]).quantize(Decimal('.01')), totalgastos = Decimal(row[13]).quantize(Decimal('.01')), excepcionesgastos = int(row[14]), valorexcepcion = Decimal(row[15]).quantize(Decimal('.01')), fraccionbasica = Decimal(row[16]).quantize(Decimal('.01')), excedentehasta = Decimal(row[17]).quantize(Decimal('.01')), impuestofraccion = Decimal(row[18]).quantize(Decimal('.01')), porcentajeimpuesto = Decimal(row[19]).quantize(Decimal('.01')), segurogastos = Decimal(row[20]).quantize(Decimal('.01')), valorretenido = Decimal(row[21]).quantize(Decimal('.01')), impuestopagar = Decimal(row[22]).quantize(Decimal('.01')), detallevivienda = Decimal(row[23]).quantize(Decimal('.01')), detalleeducacion = Decimal(row[24]).quantize(Decimal('.01')), detallesalud = Decimal(row[25]).quantize(Decimal('.01')), detallealimentacion = Decimal(row[26]).quantize(Decimal('.01')), detallevestimenta = Decimal(row[27]).quantize(Decimal('.01'))).exists():
#                 gastospersonales = GastosPersonales(persona = persona,
#                                                     periodogastospersonales = periodogastospersonales,
#                                                     mes = int(row[29]),
#                                                     rmuproyectado = Decimal(row[5]).quantize(Decimal('.01')),
#                                                     rmupagado = Decimal(row[4]).quantize(Decimal('.01')),
#                                                     horasextraspagado = Decimal(row[6]).quantize(Decimal('.01')),
#                                                     horasextrasactual = Decimal(row[7]).quantize(Decimal('.01')),
#                                                     horasextrasproyectado = Decimal(row[8]).quantize(Decimal('.01')),
#                                                     otrosingresos = Decimal(row[9]).quantize(Decimal('.01')),
#                                                     totalingresos = Decimal(row[12]).quantize(Decimal('.01')),
#                                                     otrosgastos = Decimal(row[10]).quantize(Decimal('.01')),
#                                                     rebajasotros = Decimal(row[11]).quantize(Decimal('.01')),
#                                                     totalgastos = Decimal(row[13]).quantize(Decimal('.01')),
#                                                     excepcionesgastos = int(row[14]),
#                                                     valorexcepcion = Decimal(row[15]).quantize(Decimal('.01')),
#                                                     fraccionbasica = Decimal(row[16]).quantize(Decimal('.01')),
#                                                     excedentehasta = Decimal(row[17]).quantize(Decimal('.01')),
#                                                     impuestofraccion = Decimal(row[18]).quantize(Decimal('.01')),
#                                                     porcentajeimpuesto = Decimal(row[19]).quantize(Decimal('.01')),
#                                                     segurogastos = Decimal(row[20]).quantize(Decimal('.01')),
#                                                     valorretenido = Decimal(row[21]).quantize(Decimal('.01')),
#                                                     impuestopagar = Decimal(row[22]).quantize(Decimal('.01')),
#                                                     detallevivienda = Decimal(row[23]).quantize(Decimal('.01')),
#                                                     detalleeducacion = Decimal(row[24]).quantize(Decimal('.01')),
#                                                     detallesalud = Decimal(row[25]).quantize(Decimal('.01')),
#                                                     detallealimentacion = Decimal(row[26]).quantize(Decimal('.01')),
#                                                     detallevestimenta = Decimal(row[27]).quantize(Decimal('.01')),
#                                                     retensionmensual = Decimal(row[28]).quantize(Decimal('.01')))
#                 gastospersonales.save(usuario=usuario)
#             else:
#                 gastospersonales = GastosPersonales.objects.filter(persona = persona, periodogastospersonales = periodogastospersonales, mes = int(row[29]), rmuproyectado = Decimal(row[5]).quantize(Decimal('.01')), rmupagado = Decimal(row[4]).quantize(Decimal('.01')), horasextraspagado = Decimal(row[6]).quantize(Decimal('.01')), horasextrasactual = Decimal(row[7]).quantize(Decimal('.01')), horasextrasproyectado = Decimal(row[8]).quantize(Decimal('.01')), otrosingresos = Decimal(row[9]).quantize(Decimal('.01')), totalingresos = Decimal(row[12]).quantize(Decimal('.01')), otrosgastos = Decimal(row[10]).quantize(Decimal('.01')), rebajasotros = Decimal(row[11]).quantize(Decimal('.01')), totalgastos = Decimal(row[13]).quantize(Decimal('.01')), excepcionesgastos = int(row[14]), valorexcepcion = Decimal(row[15]).quantize(Decimal('.01')), fraccionbasica = Decimal(row[16]).quantize(Decimal('.01')), excedentehasta = Decimal(row[17]).quantize(Decimal('.01')), impuestofraccion = Decimal(row[18]).quantize(Decimal('.01')), porcentajeimpuesto = Decimal(row[19]).quantize(Decimal('.01')), segurogastos = Decimal(row[20]).quantize(Decimal('.01')), valorretenido = Decimal(row[21]).quantize(Decimal('.01')), impuestopagar = Decimal(row[22]).quantize(Decimal('.01')), detallevivienda = Decimal(row[23]).quantize(Decimal('.01')), detalleeducacion = Decimal(row[24]).quantize(Decimal('.01')), detallesalud = Decimal(row[25]).quantize(Decimal('.01')), detallealimentacion = Decimal(row[26]).quantize(Decimal('.01')), detallevestimenta = Decimal(row[27]).quantize(Decimal('.01'))).order_by('-id')[0]
#
#             resumenmesgastospersonales = ResumenMesGastosPersonales(gastospersonales = gastospersonales,
#                                                                     mes = int(row[2]),
#                                                                     retensionmensual = Decimal(row[28]).quantize(Decimal('.01')))
#             resumenmesgastospersonales.save(usuario=usuario)
#         else:
#             print 'cedula %s' % row[1].strip()
#     linea += 1
#



# # alumnos del PRE nuevos
# workbook = xlrd.open_workbook("Listado de matriculados 2017 3.xlsx")
# sheet = workbook.sheet_by_index(0)
# linea = 1
#
# periodo = Periodo.objects.get(pk=14)
# for rowx in range(sheet.nrows):
#     if linea>1:
#         cols = sheet.row_values(rowx)
#         cedula = cols[0].strip().upper()
#         bandera = True
#         Persona.objects.filter(cedula=cedula).delete()
#         if Persona.objects.filter(cedula=cedula).exists():
#             bandera = False
#
#         if bandera:
#             persona = Persona(cedula=cedula,
#                               apellido1=cols[1],
#                               apellido2=cols[2],
#                               nombres=cols[3],
#                               sexo_id=int(cols[5]),
#                               # nacimiento=xlrd.xldate.xldate_as_datetime(cols[4], workbook.datemode).date(),
#                               nacimiento=convertirfecha2(cols[4]),
#                               paisnacimiento_id = int(cols[11]),
#                               provincianacimiento =None,
#                               cantonnacimiento=None,
#                               # parroquianacimiento_id=cols[8],
#                               pais_id=int(cols[12]),
#                               provincia=None,
#                               canton=None,
#                               # parroquia_id=cols[8],
#                               email=cols[17],
#                               telefono=str(cols[19]),
#                               telefono_conv =str(cols[18]))
#             persona.save()
#             username = calculate_username(persona)
#             usuario = generar_usuario(persona, username, ALUMNOS_GROUP_ID)
#             if EMAIL_INSTITUCIONAL_AUTOMATICO:
#                 persona.emailinst = username + '@' + EMAIL_DOMAIN
#                 persona.save()
#             grupo = None
#             if UTILIZA_GRUPOS_ALUMNOS:
#                 grupo = Grupo.objects.get(pk=int(cols[15]))
#                 carrera = grupo.carrera
#                 sesion = grupo.sesion
#                 modalidad = grupo.modalidad
#                 sede = grupo.sede
#             else:
#                 sesion = Sesion.objects.get(nombre=cols[7].strip())
#                 carrera = Carrera.objects.get(pk=int(cols[6]))
#                 modalidad = Modalidad.objects.get(pk=int(cols[8]))
#                 sede = Sede.objects.get(pk=1)
#         else:
#             persona = Persona.objects.filter(cedula=cols[0])[0]
#             sesion = Sesion.objects.get(nombre=cols[7].strip())
#             carrera = Carrera.objects.get(pk=int(cols[6]))
#             modalidad = Modalidad.objects.get(pk=int(cols[8]))
#             sede = Sede.objects.get(pk=1)
#
#         if not Inscripcion.objects.filter(persona=persona,carrera=carrera).exists():
#             inscripcion = Inscripcion(persona=persona,
#                                       fecha=datetime.now().date(),
#                                       carrera=carrera,
#                                       modalidad=modalidad,
#                                       sesion=sesion,
#                                       sede=sede,
#                                       colegio=cols[9])
#             inscripcion.save()
#             persona.crear_perfil(inscripcion=inscripcion)
#             documentos = DocumentosDeInscripcion(inscripcion=inscripcion,
#                                                  titulo=False,
#                                                  acta=False,
#                                                  cedula=False,
#                                                  votacion=False,
#                                                  actaconv=False,
#                                                  partida_nac=False,
#                                                  pre=False,
#                                                  observaciones_pre='',
#                                                  fotos=False)
#             documentos.save()
#             preguntasinscripcion = inscripcion.preguntas_inscripcion()
#             perfil_inscripcion = inscripcion.persona.mi_perfil()
#
#             perfil_inscripcion.raza_id =  int(cols[16])
#             perfil_inscripcion.save()
#             inscripciontesdrive = InscripcionTesDrive(inscripcion=inscripcion,
#                                                       licencia=False,
#                                                       record=False,
#                                                       certificado_tipo_sangre=False,
#                                                       prueba_psicosensometrica=False,
#                                                       certificado_estudios=False)
#             inscripciontesdrive.save()
#             # inscripcion.mi_malla()
#             inscripcion.malla_inscripcion()
#             # inscripcion.actualizar_nivel()
#             if USA_TIPOS_INSCRIPCIONES:
#                 inscripciontipoinscripcion = InscripcionTipoInscripcion(inscripcion=inscripcion,
#                                                                         tipoinscripcion_id=TIPO_INSCRIPCION_INICIAL)
#                 inscripciontipoinscripcion.save()
#             # persona.creacion_persona(request.session['nombresistema'])
#         else:
#             inscripcion = Inscripcion.objects.filter(persona=persona,carrera=carrera)[0]
#             perfil_inscripcion = inscripcion.persona.mi_perfil()
#             perfil_inscripcion.raza_id = int(cols[16])
#             perfil_inscripcion.save()
#
#         nivel = Nivel.objects.get(periodo=periodo, sesion=sesion, paralelo__icontains='ADMI')
#         # matricula
#         if not inscripcion.matricula_periodo(periodo):
#             matricula = Matricula(inscripcion=inscripcion,
#                                   nivel=nivel,
#                                   pago=False,
#                                   iece=False,
#                                   becado=False,
#                                   porcientobeca=0,
#                                   fecha=datetime.now().date(),
#                                   hora=datetime.now().time(),
#                                   fechatope=fechatope(datetime.now().date()))
#             matricula.save()
#         else:
#             matricula = Matricula.objects.get(inscripcion=inscripcion, nivel=nivel)
#         for materia in Materia.objects.filter(nivel__periodo=periodo, paralelo=cols[22].strip(), asignaturamalla__malla__carrera=carrera, nivel__sesion=sesion):
#             if not MateriaAsignada.objects.filter(matricula=matricula,materia=materia).exists():
#                 matriculas = matricula.inscripcion.historicorecordacademico_set.filter(asignatura=materia.asignatura, fecha__lt=materia.nivel.fin).count() + 1
#                 materiaasignada = MateriaAsignada(matricula=matricula,
#                                                   materia=materia,
#                                                   notafinal=0,
#                                                   asistenciafinal=0,
#                                                   cerrado=False,
#                                                   matriculas=matriculas,
#                                                   observaciones='',
#                                                   estado_id=NOTA_ESTADO_EN_CURSO)
#                 materiaasignada.save()
#                 materiaasignada.asistencias()
#                 materiaasignada.evaluacion()
#                 materiaasignada.mis_planificaciones()
#                 materiaasignada.save()
#         matricula.actualizar_horas_creditos()
#         matricula.estado_matricula=2
#         matricula.save()
#
#     linea += 1
#     print(linea)

# # eliminar rol detalle
# workbook = xlrd.open_workbook("eliminarrol.xlsx")
# sheet = workbook.sheet_by_index(0)
# linea = 1
# for rowx in range(sheet.nrows):
#     if linea > 1:
#         print linea
#         cols = sheet.row_values(rowx)
#         persona = None
#         if Persona.objects.filter(cedula=cols[1].strip()).exists():
#             persona = Persona.objects.filter(cedula=cols[1].strip())[0]
#
#         if persona:
#             if DetallePeriodoRol.objects.filter(persona=persona, periodo__id=int(cols[0].strip()), rubro__id=int(cols[2].strip()), valor=Decimal(cols[3].strip()).quantize(Decimal('.01'))).exists():
#                 DetallePeriodoRol.objects.filter(persona=persona, periodo__id=int(cols[0].strip()), rubro__id=int(cols[2].strip()), valor=Decimal(cols[3].strip()).quantize(Decimal('.01'))).update(status=False)
#         else:
#             print 'cedula: %s' %  cols[1].strip()
#     linea += 1
#
# # ERROR LOYOLA
# workbook = xlrd.open_workbook("error.xls")
# sheet = workbook.sheet_by_index(0)
# linea = 1
# for rowx in range(sheet.nrows):
#     if linea>1:
#         cols = sheet.row_values(rowx)
#         if Inscripcion.objects.filter(persona__cedula=cols[0].strip(), carrera__id=int(cols[1]), fecha=convertirfecha2('2016-10-07')).exists():
#             if Inscripcion.objects.filter(persona__cedula=cols[0].strip(), fecha=convertirfecha2('2016-10-07')).exclude(carrera__id=int(cols[1])).exists():
#                 # inscripcion2.delete()
#                 inscripcion2 = Inscripcion.objects.filter(persona__cedula=cols[0].strip(), fecha=convertirfecha2('2016-10-07')).exclude(carrera__id=int(cols[1]))[0]
#                 print '%s - %s' % (inscripcion2, inscripcion2.carrera)
#                 inscripcion2.delete()
#         else:
#             data = [[cols[0].strip(), "I"]]
#             dataWriter.writerows(data)
#     linea += 1


# # formulario 107
# workbook = xlrd.open_workbook("107.xlsx")
# sheet = workbook.sheet_by_index(0)
# linea = 1
# for rowx in range(sheet.nrows):
#     print linea
#     if linea>1:
#         cols = sheet.row_values(rowx)
#
#         persona = None
#         if Persona.objects.filter(cedula=cols[0].strip()).exists():
#             persona = Persona.objects.filter(cedula=cols[0].strip())[0]
#
#         periodo = None
#         if PeriodoGastosPersonales.objects.filter(anio=int(cols[1].strip())).exists():
#             periodo = PeriodoGastosPersonales.objects.filter(anio=int(cols[1].strip()))[0]
#
#         excepcionesgastos = 0
#         if int(Decimal(cols[15].strip()).quantize(Decimal('.01'))) == 0 and int(Decimal(cols[16].strip()).quantize(Decimal('.01'))) == 0:
#             excepcionesgastos = 1
#         if int(Decimal(cols[15].strip()).quantize(Decimal('.01'))) > 0 and int(Decimal(cols[16].strip()).quantize(Decimal('.01'))) == 0:
#             excepcionesgastos = 3
#         if int(Decimal(cols[15].strip()).quantize(Decimal('.01'))) == 0 and int(Decimal(cols[16].strip()).quantize(Decimal('.01'))) > 0:
#             excepcionesgastos = 2
#         if int(Decimal(cols[15].strip()).quantize(Decimal('.01'))) > 0 and int(Decimal(cols[16].strip()).quantize(Decimal('.01'))) > 0:
#             excepcionesgastos = 4
#
#
#         if persona:
#             formulario = DeclaracionSriAnual (id = int(cols[29].strip()),
#                                               persona = persona,
#                                               periodogastospersonales = periodo,
#                                               sueldosysalarios = Decimal(cols[2].strip()).quantize(Decimal('.01')),
#                                               sobresueldos = Decimal(cols[3].strip()).quantize(Decimal('.01')),
#                                               participacionutil = Decimal(cols[4].strip()).quantize(Decimal('.01')),
#                                               otrosingresos = Decimal(cols[5].strip()).quantize(Decimal('.01')),
#                                               decimotercer = Decimal(cols[6].strip()).quantize(Decimal('.01')),
#                                               decimocuarto = Decimal(cols[7].strip()).quantize(Decimal('.01')),
#                                               fondoreserva = Decimal(cols[8].strip()).quantize(Decimal('.01')),
#                                               otrosingresossinrenta = Decimal(cols[9].strip()).quantize(Decimal('.01')),
#                                               fraccionbasica = Decimal(cols[10].strip()).quantize(Decimal('.01')),
#                                               aportepersonaleste = Decimal(cols[11].strip()).quantize(Decimal('.01')),
#                                               totalingresos = Decimal(cols[12].strip()).quantize(Decimal('.01')),
#                                               otrosgastos = Decimal(cols[13].strip()).quantize(Decimal('.01')),
#                                               totalgastos = Decimal(cols[21].strip()).quantize(Decimal('.01'))+Decimal(cols[22].strip()).quantize(Decimal('.01'))+Decimal(cols[23].strip()).quantize(Decimal('.01'))+Decimal(cols[24].strip()).quantize(Decimal('.01'))+Decimal(cols[25].strip()).quantize(Decimal('.01')),
#                                               excepcionesgastos = excepcionesgastos,
#                                               valorexcepcion = Decimal(cols[15].strip()).quantize(Decimal('.01'))+Decimal(cols[16].strip()).quantize(Decimal('.01')),
#                                               segurogastos = Decimal(cols[17].strip()).quantize(Decimal('.01')),
#                                               valorretenido = Decimal(cols[18].strip()).quantize(Decimal('.01')),
#                                               impuestoasumido = Decimal(cols[19].strip()).quantize(Decimal('.01')),
#                                               impuestopagar = Decimal(cols[20].strip()).quantize(Decimal('.01')),
#                                               detallevivienda = Decimal(cols[21].strip()).quantize(Decimal('.01')),
#                                               detalleeducacion = Decimal(cols[22].strip()).quantize(Decimal('.01')),
#                                               detallesalud = Decimal(cols[23].strip()).quantize(Decimal('.01')),
#                                               detallealimentacion = Decimal(cols[24].strip()).quantize(Decimal('.01')),
#                                               detallevestimenta = Decimal(cols[25].strip()).quantize(Decimal('.01')),
#                                               retensionmensual = Decimal(cols[26].strip()).quantize(Decimal('.01')),
#                                               baseimponible = Decimal(cols[2].strip()).quantize(Decimal('.01')) + Decimal(cols[3].strip()).quantize(Decimal('.01')) + Decimal(cols[5].strip()).quantize(Decimal('.01')) - Decimal(cols[13].strip()).quantize(Decimal('.01')) - (Decimal(cols[15].strip()).quantize(Decimal('.01'))+Decimal(cols[16].strip()).quantize(Decimal('.01'))) - (Decimal(cols[21].strip()).quantize(Decimal('.01'))+Decimal(cols[22].strip()).quantize(Decimal('.01'))+Decimal(cols[23].strip()).quantize(Decimal('.01'))+Decimal(cols[24].strip()).quantize(Decimal('.01'))+Decimal(cols[25].strip()).quantize(Decimal('.01'))) - Decimal(cols[11].strip()).quantize(Decimal('.01')),
#                                               ingresosgravados = Decimal(cols[2].strip()).quantize(Decimal('.01')) + Decimal(cols[3].strip()).quantize(Decimal('.01')))
#
#             formulario.save()
#         else:
#             print 'cedula no existe %s' % cols[0].strip()
#
#     linea += 1

# # eliminar matricula
# workbook = xlrd.open_workbook("eliminarmatricula.xlsx")
# sheet = workbook.sheet_by_index(0)
# linea = 1
# for rowx in range(sheet.nrows):
#     if linea>1:
#         cols = sheet.row_values(rowx)
#         if Matricula.objects.filter(inscripcion__id=int(cols[0]), nivel__periodo__id=9).exists():
#             matricula = Matricula.objects.filter(inscripcion__id=int(cols[0]), nivel__periodo__id=9)[0]
#             for materiaasignada in matricula.materiaasignada_set.all():
#                 materiaasignada.delete()
#                 print materiaasignada
#             matricula.delete()
#             print matricula
#         else:
#             print 'no esta matriculado : %s' % int(cols[0])
#     linea += 1

# eliminar materias del record

# # eliminar matricula
# workbook = xlrd.open_workbook("eliminarrecord.xlsx")
# sheet = workbook.sheet_by_index(0)
# linea = 1
# for rowx in range(sheet.nrows):
#     if linea>1:
#         cols = sheet.row_values(rowx)
#         if RecordAcademico.objects.filter(inscripcion__id=int(cols[0]), asignatura__id=int(cols[1])).exists():
#             record = RecordAcademico.objects.get(inscripcion__id=int(cols[0]), asignatura__id=int(cols[1]))
#             asignatura = record.asignatura
#             inscripcion = record.inscripcion
#             homologacion = record.homologacioninscripcion_set.all()
#             homologacion.delete()
#             convalidacion = record.convalidacioninscripcion_set.all()
#             convalidacion.delete()
#             historico = record.historicorecordacademico_set.all()
#             historico.delete()
#             record.delete()
#             inscripcion.actualizar_nivel()
#             inscripcion.actualiza_matriculas(asignatura)
#             # print 'inscripcion %s asignatura %s' % (record.inscripcion,record)
#         else:
#             print 'inscripcion %s asignatura %s' % (cols[0],cols[1])
#
#     linea += 1


# factura = Factura.objects.filter(pk=183)[0]
# factura.pagos.update(comprobante=factura.comprobante)
# for pago in Pago.objects.filter(comprobante=factura.comprobante):
#     saldopartida = pago.rubro.tipo.partida_saldo(pago.fecha.year)
#     if not saldopartida:
#         print 'error'
#     detallecomprobante = ResumenComprobantePartida(comprobanterecaudacion=factura.comprobante,
#                                                    partida=saldopartida.partidassaldo,
#                                                    valor=pago.valortotal)
#     detallecomprobante.save()

# # lenar tabla de faltas asistencias
# fechas = []
# inicio=convertirfecha2('2017-10-03')
# fin=convertirfecha2('2018-01-18')
# FaltasMateriaPeriodo.objects.filter(periodo__id=14).delete()
# for dia in daterange(inicio, (fin + timedelta(days=1))):
#     fechas.append(dia)
#
# for fechatope in fechas:
#     print (fechatope)
#     # fechatope='2017-01-31'
#     inicio_aux=''
#     fin_aux=''
#
#     periodos=Periodo.objects.filter(inicio__lte=fechatope, fin__gte=fechatope)
#     for periodo in periodos:
#         profesormaterias=ProfesorMateria.objects.filter(materia__nivel__periodo=periodo)
#         diasnolaborables = DiasNoLaborable.objects.filter(periodo=periodo).order_by('fecha')
#         claseshorarios = Clase.objects.filter(materia__nivel__periodo=periodo, activo=True).distinct().order_by('materia__profesormateria__profesor', 'turno__comienza')
#         n = 1
#         for profesormateria in ProfesorMateria.objects.filter(materia__nivel__periodo=periodo):
#             print("%s - %s" % (n,fechatope))
#             n += 1
#             profesorid = profesormateria.profesor.id
#             claseshorario = claseshorarios.filter(materia__profesormateria__profesor__id=profesorid)
#
#             inicio = profesormateria.materia.inicio
#             fin = profesormateria.materia.fin
#             fecha = convertirfecha2(fechatope)
#             if fin > fecha:
#                 fin = fecha
#
#             if inicio <= fin:
#                 if not (inicio == inicio_aux and fin == fin_aux):
#                     fechas = []
#                     if inicio == fin:
#                         fechas.append(inicio)
#                         dia_semana = inicio.isoweekday()
#                         claseshorario = claseshorario.filter(dia=dia_semana)
#                     else:
#                         for dia in daterange(inicio, (fin + timedelta(days=1))):
#                             fechas.append(dia)
#                 profesormateria.faltas_docente(periodo, inicio, fin, diasnolaborables, claseshorario, fechas)




# # cron de falta asistencia
# fechatope = datetime.now().date() - timedelta(days=1)
# periodos=Periodo.objects.filter(inicio__lte=fechatope, fin__gte=fechatope)
# for periodo in periodos:
#     if DiasNoLaborable.objects.filter(periodo=periodo, fecha=fechatope).exists():
#         diasnolaborables = DiasNoLaborable.objects.filter(periodo=periodo, fecha=fechatope).order_by('fecha')[0]
#         coordinacion = None
#         if diasnolaborables.coordinacion:
#             coordinacion = diasnolaborables.coordinacion
#
#         carrera = None
#         if diasnolaborables.carrera:
#             carrera = diasnolaborables.carrera
#
#         nivelmalla = None
#         if diasnolaborables.nivelmalla:
#             nivelmalla = diasnolaborables.nivelmalla
#         else:
#             if coordinacion and carrera and nivelmalla:
#                 FaltasMateriaPeriodo.objects.filter(periodo=periodo, materia__asignaturamalla__malla__carrera__coordinacion_carrera__coordinacion=coordinacion, materia__asignaturamalla__malla__carrera=carrera, materia__asignaturamalla__nivelmalla=nivelmalla, fecha=fechatope).delete()
#             else:
#                 if coordinacion and carrera:
#                     FaltasMateriaPeriodo.objects.filter(periodo=periodo, materia__asignaturamalla__malla__carrera__coordinacion_carrera__coordinacion=coordinacion, materia__asignaturamalla__malla__carrera=carrera, fecha=fechatope).delete()
#                 else:
#                     if coordinacion:
#                         FaltasMateriaPeriodo.objects.filter(periodo=periodo, materia__asignaturamalla__malla__carrera__coordinacion_carrera__coordinacion=coordinacion, fecha=fechatope).delete()
#                     else:
#                         FaltasMateriaPeriodo.objects.filter(periodo=periodo, fecha=fechatope).delete()
#
#     inicio_aux=''
#     fin_aux=''
#     profesormaterias=ProfesorMateria.objects.filter(materia__nivel__periodo=periodo)
#     diasnolaborables = DiasNoLaborable.objects.filter(periodo=periodo, fecha=fechatope).order_by('fecha')
#     claseshorarios = Clase.objects.filter(materia__nivel__periodo=periodo, activo=True).distinct().order_by('materia__profesormateria__profesor', 'turno__comienza')
#     n = 1
#     for profesormateria in ProfesorMateria.objects.filter(materia__nivel__periodo=periodo, desde__lte=fechatope, hasta__gte=fechatope):
#         # print n
#         n += 1
#         profesorid = profesormateria.profesor.id
#         claseshorario = claseshorarios.filter(materia__profesormateria__profesor__id=profesorid)
#
#         inicio = profesormateria.materia.inicio
#         fin = profesormateria.materia.fin
#         fecha = fechatope
#         bandera = False
#         if inicio <= fecha:
#             if fin >= fecha:
#                 bandera = True
#                 inicio = fecha
#                 fin = fecha
#
#         if bandera:
#             if not (inicio == inicio_aux and fin == fin_aux):
#                 fechas = []
#                 if inicio == fin:
#                     fechas.append(inicio)
#                     dia_semana = inicio.isoweekday()
#                     claseshorario = claseshorario.filter(dia=dia_semana)
#                 else:
#                     for dia in daterange(fechatope, (fechatope + timedelta(days=1))):
#                         fechas.append(dia)
#             profesormateria.faltas_docente(periodo, inicio, fin, diasnolaborables, claseshorario, fechas)




# materias = Materia.objects.filter(nivel__periodo_id=10).order_by('-id')
# for materia in materias:
#     print materia.id
#     evaluaciones = EvaluacionGenerica.objects.filter(materiaasignada__materia=materia)
#     evaluaciones.delete()
#     for maa in materia.asignados_a_esta_materia():
#         maa.evaluacion()
#         maa.notafinal = 0
#         maa.save()
#     if materia.cronogramaevaluacionmodelo_set.exists():
#         cronograma = materia.cronogramaevaluacionmodelo_set.all()[0]
#         cronograma.materias.remove(materia)
# print 'listo'


# # actualizar promedio notas
# # ,materiaasignada__materia__id=6797, materiaasignada__id=159063
# for materiaasignadaplani in MateriaAsignadaPlanificacion.objects.filter(materiaasignada__matricula__nivel__periodo__id=10).order_by('materiaasignada__id').distinct():
#     print materiaasignadaplani.materiaasignada.id
#     modeloevaluativo = materiaasignadaplani.materiaasignada.materia.modeloevaluativo
#     materiaasignada_id = materiaasignadaplani.materiaasignada.id
#     valor = materiaasignadaplani.promedio_calificacion()
#     sel_id = materiaasignadaplani.planificacion.tipoevaluacion.nombre
#     actualizar_nota_planificacion(materiaasignada_id, sel_id, valor)
#     # campo = materiaasignadaplani.materiaasignada.campo(materiaasignadaplani.planificacion.tipoevaluacion.nombre)
#     # campo.valor = valor
#     # campo.save()
# print 'listo'


# # Actualizacion Catalogo
# workbook = xlrd.open_workbook("catalogo.xlsx")
# sheet = workbook.sheet_by_index(0)
# linea = 1
# for rowx in range(sheet.nrows):
#     if linea>1:
#         cols = sheet.row_values(rowx)
#         if Partida.objects.filter(codigo=cols[0].strip()).exists():
#             partida = Partida.objects.filter(codigo=cols[0].strip())[0]
#             if CatalogoBien.objects.filter(identificador=cols[1].strip()).exists():
#                 CatalogoBien.objects.filter(identificador=cols[1].strip()).update(item=partida)
#             else:
#                 print 'No existe catalago %s' % (cols[1])
#         else:
#             print 'No existe partida %s' % (cols[0])
#
#     linea += 1
#     print linea
# print 'listo'


# actualizo responsable activo
# a = TraspasoActivo.objects.filter(tipo=2, custodiobienrecibe__isnull=False, usuariobienrecibe__isnull=True).count()
# print a


# for a in ActivoFijo.objects.filter(status=True, statusactivo=1, responsable__isnull=True):
#     if TraspasoActivo.objects.filter(detalletraspasoactivo__activo=a, detalletraspasoactivo__seleccionado=True, usuariobienrecibe__isnull=False, estado=2, tipo=2).exists():
#         usuario = TraspasoActivo.objects.filter(detalletraspasoactivo__activo=a, detalletraspasoactivo__seleccionado=True,usuariobienrecibe__isnull=False, estado=2, tipo=2).order_by("-numero").distinct()[0].usuariobienrecibe
#         custodio = TraspasoActivo.objects.filter(detalletraspasoactivo__activo=a, detalletraspasoactivo__seleccionado=True,  custodiobienrecibe__isnull=False, estado=2, tipo=2).order_by("-numero").distinct()[0].custodiobienrecibe
#         ubicacion = TraspasoActivo.objects.filter(detalletraspasoactivo__activo=a, detalletraspasoactivo__seleccionado=True, ubicacionbienrecibe__isnull=False, estado=2, tipo=2).order_by("-numero").distinct()[0].ubicacionbienrecibe
#         a.custodio = custodio
#         a.ubicacion = ubicacion
#         a.responsable = usuario
#     elif TraspasoActivo.objects.filter(detalletraspasoactivo__activo=a, detalletraspasoactivo__seleccionado=True, usuariobienrecibe__isnull=False, estado=2, tipo=1).exists():
#         usuario = \
#         TraspasoActivo.objects.filter(detalletraspasoactivo__activo=a, detalletraspasoactivo__seleccionado=True,
#                                       usuariobienrecibe__isnull=False, estado=2, tipo=1).order_by("-numero").distinct()[0].usuariobienrecibe
#         a.responsable = usuario
#     else:
#         a.custodio = None
#         a.ubicacion = None
#         a.responsable = None
#     a.save()
#     print a.responsable


# for comprobante in ComprobanteRecaudacion.objects.filter(tipocomprobanterecaudacion__id__in=[1,2,3,7]).order_by('numero'):
#     comprobante.resumencomprobantepartida_set.all().delete()
#     for pago in Pago.objects.filter(comprobante=comprobante):
#         saldopartida = pago.rubro.tipo.partida_saldo(pago.fecha.year)
#         if not saldopartida:
#             print pago.rubro.tipo
#             pass
#         if ResumenComprobantePartida.objects.filter(comprobanterecaudacion=comprobante,
#                                                     partida=saldopartida.partidassaldo).exists():
#             detalle = ResumenComprobantePartida.objects.filter(comprobanterecaudacion=comprobante,
#                                                                partida=saldopartida.partidassaldo)[0]
#             detalle.valor += pago.valortotal
#             detalle.save()
#         else:
#             detallecomprobante = ResumenComprobantePartida(comprobanterecaudacion=comprobante,
#                                                            partida=saldopartida.partidassaldo,
#                                                            valor=pago.valortotal)
#             detallecomprobante.save()
#     print comprobante.numero


# for pac in Pac.objects.all():
#     pac.save()
#     print pac.id
# print 'listo'

# traspaso = TraspasoActivo.objects.get(tipo=2, numero=2226)
# for detalle in traspaso.detalletraspasoactivo_set.all():
#     if TraspasoActivo.objects.filter(detalletraspasoactivo__activo=detalle.activo).exists():
#         usuario = None
#         if TraspasoActivo.objects.filter(detalletraspasoactivo__activo=detalle.activo, detalletraspasoactivo__seleccionado=True, usuariobienrecibe__isnull=False, estado=2, tipo=2).exists():
#             usuario = TraspasoActivo.objects.filter(detalletraspasoactivo__activo=detalle.activo,
#                                                     detalletraspasoactivo__seleccionado=True,
#                                                     usuariobienrecibe__isnull=False, estado=2, tipo=2).order_by("-fecha").distinct()[0].usuariobienrecibe
#
#         else:
#             usuario = TraspasoActivo.objects.filter(detalletraspasoactivo__activo=detalle.activo,
#                                                     detalletraspasoactivo__seleccionado=True,
#                                                     usuariobienrecibe__isnull=False, estado=2, tipo=1).order_by("-fecha").distinct()[0].usuariobienrecibe
#         custodio = TraspasoActivo.objects.filter(detalletraspasoactivo__activo=detalle.activo, detalletraspasoactivo__seleccionado=True, custodiobienrecibe__isnull=False, estado=2).order_by("-fecha").distinct()[0].custodiobienrecibe
#         ubicacion = TraspasoActivo.objects.filter(detalletraspasoactivo__activo=detalle.activo, detalletraspasoactivo__seleccionado=True, ubicacionbienrecibe__isnull=False, estado=2).order_by("-fecha").distinct()[0].ubicacionbienrecibe
#         detalle.activo.custodio = custodio
#         detalle.activo.ubicacion = ubicacion
#         detalle.activo.responsable = usuario
#     else:
#         detalle.activo.custodio = None
#         detalle.activo.ubicacion = None
#         detalle.activo.responsable = None
#     detalle.activo.save()


# # pre
# periodo = Periodo.objects.filter(pk=10)[0]
# personas = Persona.objects.filter(inscripcion__matricula__nivel__periodo=periodo, usuario__isnull=True)
# for persona in personas:
#     print persona
#     username = calculate_username(persona)
#     generar_usuario(persona, username, ALUMNOS_GROUP_ID)
#     if EMAIL_INSTITUCIONAL_AUTOMATICO:
#         persona.emailinst = username + '@' + EMAIL_DOMAIN
#         persona.save()



# # descargar foto
# cursor = connection.cursor()
# sql = "SELECT p.cedula, p.apellido1||' '||p.apellido2||' '||p.nombres as nombre, (CASE p.sexo_id WHEN 1 THEN 'FEMENINO' ELSE 'MASCULINO' END) AS sexo, " \
#       " COALESCE((select fp.foto from sga_fotopersona fp where fp.persona_id=p.id and fp.status=true),'') as direccion_foto " \
#       " from sga_persona p, auth_user usua, sagest_distributivopersona d, sagest_nivelocupacional nio, sga_profesor pro " \
#       " where p.id=d.persona_id and d.status=true and d.estadopuesto_id=1 and d.regimenlaboral_id=2 and pro.persona_id=p.id and pro.activo=true " \
#       " and usua.id=p.usuario_id and d.nivelocupacional_id=nio.id order by p.apellido1,p.apellido2"
#
# cursor.execute(sql)
# results = cursor.fetchall()
# for per in results:
#     direccion = 'http://sga.unemi.edu.ec/media/'+per[3]
#     print 'foto %s' %  direccion
#     print 'cedula %s' % per[0]
#     download(direccion, per[0]+'-'+per[1])
# print 'listo'


# # Colegio
# workbook = xlrd.open_workbook("subircolegios.xlsx")
# sheet = workbook.sheet_by_index(0)
# linea = 1
# for rowx in range(sheet.nrows):
#     if linea>1:
#         cols = sheet.row_values(rowx)
#
#         if Colegio.objects.filter(nombre=cols[2].strip(), cantond = cols[1].strip(), tipo = int(cols[3])).exists():
#             canton = None
#             if Canton.objects.filter(nombre=cols[1].strip()).count() == 1:
#                 canton = Canton.objects.filter(nombre=cols[1].strip())[0]
#             else:
#                 provincia = None
#                 if Provincia.objects.filter(nombre=cols[0].strip()).count() == 1:
#                     provincia = Provincia.objects.filter(nombre=cols[0].strip())[0]
#                     if Canton.objects.filter(nombre=cols[1].strip(), provincia=provincia).exists():
#                         canton = Canton.objects.filter(nombre=cols[1].strip(), provincia=provincia)[0]
#
#             colegio = Colegio.objects.filter(nombre=cols[2].strip())[0]
#             colegio.cantond = cols[1].strip()
#             colegio.tipo = int(cols[3])
#             colegio.canton = canton
#             colegio.save()
#         else:
#             canton = None
#             if Canton.objects.filter(nombre=cols[1].strip()).count() == 1:
#                 canton = Canton.objects.filter(nombre=cols[1].strip())[0]
#             else:
#                 provincia = None
#                 if Provincia.objects.filter(nombre=cols[0].strip()).count() == 1:
#                     provincia = Provincia.objects.filter(nombre=cols[0].strip())[0]
#                     if Canton.objects.filter(nombre=cols[1].strip(), provincia=provincia).exists():
#                         canton = Canton.objects.filter(nombre=cols[1].strip(), provincia=provincia)[0]
#
#             colegio = Colegio(nombre = cols[2].strip(),
#                               cantond = cols[1].strip(),
#                               canton = canton,
#                               tipo = int(cols[3]))
#             colegio.save()
#     linea += 1
#     print linea
# print 'listo'

# formulario 107
# workbook = xlrd.open_workbook("107a.xlsx")
# sheet = workbook.sheet_by_index(1)
# linea = 1
# periodo = PeriodoGastosPersonales.objects.filter(anio=2016)[0]
# DeclaracionSriAnual.objects.filter(periodogastospersonales=periodo).delete()
# for rowx in range(sheet.nrows):
#     if linea>1:
#         cols = sheet.row_values(rowx)
#         print cols[2].strip()
#         persona = Persona.objects.filter(Q(cedula=cols[2].strip()) | Q(pasaporte=cols[2].strip()))[0]
#         gasto = None
#         if GastosPersonales.objects.filter(persona=persona, periodogastospersonales=periodo).exists():
#             gasto = GastosPersonales.objects.filter(persona=persona, periodogastospersonales=periodo)[0]
#         excepcionesgastos = 0
#         if (Decimal(cols[32]).quantize(Decimal('.01'))==0 and Decimal(cols[33]).quantize(Decimal('.01'))==0):
#             excepcionesgastos = 1
#         else:
#             if (Decimal(cols[32]).quantize(Decimal('.01')) > 0 and Decimal(cols[33]).quantize(Decimal('.01')) == 0):
#                 excepcionesgastos = 3
#             else:
#                 if (Decimal(cols[32]).quantize(Decimal('.01')) == 0 and Decimal(cols[33]).quantize(Decimal('.01')) > 0):
#                     excepcionesgastos = 2
#                 else:
#                     excepcionesgastos = 4
#         fraccionbasica = 0
#         if gasto:
#             fraccionbasica = gasto.fraccionbasica
#
#         declaracion = DeclaracionSriAnual(persona = persona,
#                                           periodogastospersonales = periodo,
#                                           sueldosysalarios = Decimal(cols[13]).quantize(Decimal('.01')),
#                                           sobresueldos = Decimal(cols[14]).quantize(Decimal('.01')),
#                                           participacionutil = Decimal(cols[15]).quantize(Decimal('.01')),
#                                           otrosingresos = Decimal(cols[22]).quantize(Decimal('.01')),
#                                           decimotercer = Decimal(cols[18]).quantize(Decimal('.01')),
#                                           decimocuarto = Decimal(cols[19]).quantize(Decimal('.01')),
#                                           fondoreserva = Decimal(cols[20]).quantize(Decimal('.01')),
#                                           otrosingresossinrenta = 0,
#                                           fraccionbasica = fraccionbasica,
#                                           aportepersonaleste = Decimal(cols[25]).quantize(Decimal('.01')),
#                                           totalingresos = (Decimal(cols[16]).quantize(Decimal('.01'))+Decimal(cols[23]).quantize(Decimal('.01'))),
#                                           otrosgastos = 0,
#                                           totalgastos = (Decimal(cols[27]).quantize(Decimal('.01'))+Decimal(cols[28]).quantize(Decimal('.01'))+Decimal(cols[29]).quantize(Decimal('.01'))+Decimal(cols[30]).quantize(Decimal('.01'))+Decimal(cols[31]).quantize(Decimal('.01'))),
#                                           excepcionesgastos = excepcionesgastos,
#                                           valorexcepcionedad = Decimal(cols[33]).quantize(Decimal('.01')),
#                                           valorexcepciondiscapacidad = Decimal(cols[32]).quantize(Decimal('.01')),
#                                           # segurogastos = Decimal(cols[13]).quantize(Decimal('.01')),
#                                           valorretenido = Decimal(cols[38]).quantize(Decimal('.01')),
#                                           impuestoasumido = Decimal(cols[37]).quantize(Decimal('.01')),
#                                           impuestopagar = Decimal(cols[35]).quantize(Decimal('.01')),
#                                           detallevivienda = Decimal(cols[27]).quantize(Decimal('.01')),
#                                           detalleeducacion = Decimal(cols[29]).quantize(Decimal('.01')),
#                                           detallesalud = Decimal(cols[28]).quantize(Decimal('.01')),
#                                           detallealimentacion = Decimal(cols[30]).quantize(Decimal('.01')),
#                                           detallevestimenta = Decimal(cols[31]).quantize(Decimal('.01')),
#                                           retensionmensual = Decimal(cols[13]).quantize(Decimal('.01')),
#                                           baseimponible = Decimal(cols[34]).quantize(Decimal('.01')),
#                                           ingresosgravados =Decimal(cols[23]).quantize(Decimal('.01'))                                          )
#         declaracion.save()
#         declaracion.segurogastos = declaracion.totalingresos * Decimal(PORCENTAJE_SEGURO).quantize(Decimal('.01'))
#         declaracion.save()
#     linea += 1





# # formulario 107 segunda parte
# workbook = xlrd.open_workbook("107a.xlsx")
# sheet = workbook.sheet_by_index(1)
# linea = 1
# periodo = PeriodoGastosPersonales.objects.filter(anio=2016)[0]
# DeclaracionSriAnual107A.objects.filter(periodogastospersonales=periodo).delete()
# for rowx in range(sheet.nrows):
#     if linea>1:
#         cols = sheet.row_values(rowx)
#         print cols[2].strip()
#         persona = Persona.objects.filter(Q(cedula=cols[2].strip()) | Q(pasaporte=cols[2].strip()))[0]
#
#         declaracionsrianual107a = DeclaracionSriAnual107A(persona = persona,
#                                                           periodogastospersonales = periodo,
#                                                           remuneracioncontribucionconeste = Decimal(cols[39]).quantize(Decimal('.01')),
#                                                           remuneracioncontribucionconotro = Decimal(cols[40]).quantize(Decimal('.01')),
#                                                           exoneracionremuneracion = Decimal(cols[41]).quantize(Decimal('.01')),
#                                                           totalremuneracion = Decimal(cols[42]).quantize(Decimal('.01')),
#                                                           numeromesestrabajadosconeste = Decimal(cols[43]).quantize(Decimal('.01')),
#                                                           numeromesestrabajadosconotro = Decimal(cols[44]).quantize(Decimal('.01')),
#                                                           totalnumeromeses = Decimal(cols[45]).quantize(Decimal('.01')),
#                                                           remuneracionmensualpromedio = Decimal(cols[46]).quantize(Decimal('.01')),
#                                                           numeromesescontribucionconeste = Decimal(cols[47]).quantize(Decimal('.01')),
#                                                           numeromesescontribucionconotro = Decimal(cols[48]).quantize(Decimal('.01')),
#                                                           totalnumeromesescontribucion = Decimal(cols[49]).quantize(Decimal('.01')),
#                                                           totalcontribuciongenerada = Decimal(cols[50]).quantize(Decimal('.01')),
#                                                           creditotributariodonacioncontribucionotro = Decimal(cols[51]).quantize(Decimal('.01')),
#                                                           creditotributariodonacioncontribucioneste = Decimal(cols[52]).quantize(Decimal('.01')),
#                                                           creditotributariopordonacioncontribucionnoutilizadoporeste = Decimal(cols[53]).quantize(Decimal('.01')),
#                                                           totalcreditotributariodonacioncontribucion = Decimal(cols[54]).quantize(Decimal('.01')),
#                                                           contribucionpagar = Decimal(cols[55]).quantize(Decimal('.01')),
#                                                           contribucionasumidaotro = Decimal(cols[56]).quantize(Decimal('.01')),
#                                                           contribucionretenidaotro = Decimal(cols[57]).quantize(Decimal('.01')),
#                                                           contribucionasumidaeste = Decimal(cols[58]).quantize(Decimal('.01')),
#                                                           contribucionretenidaeste = Decimal(cols[59]).quantize(Decimal('.01')))
#         declaracionsrianual107a.save()
#     linea += 1
#
# cerrar materias
# ,asignaturamalla__malla__carrera__in=[22,24,34]
# for materia in Materia.objects.filter(nivel__periodo__id__in=[14], cerrado=False).order_by('-id'):
#     if materia.cerrado==False:
#         materia.cerrado = True
#         materia.fechacierre = datetime.now().date()
#         materia.save()
#         for asig in materia.asignados_a_esta_materia():
#             asig.cerrado = True
#             asig.save()
#             asig.actualiza_estado()
#             asig.cierre_materia_asignada()
#             # for asig in materia.asignados_a_esta_materia():
#
#         print(materia.id)
# contador = 0
# # for evidencia in PracticasPreprofesionalesInscripcion.objects.filter(fechadesde__gte=datetime.strptime('01-06-2018', "%d-%m-%Y").date(), detalleevidenciaspracticaspro__estadotutor=0, detalleevidenciaspracticaspro__archivo__isnull=False, detalleevidenciaspracticaspro__estadorevision=1, detalleevidenciaspracticaspro__evidencia__id__in=[9]):
# for evidencia in DetalleEvidenciasPracticasPro.objects.filter(inscripcionpracticas__fechadesde__gte=datetime.strptime('01-06-2018', "%d-%m-%Y").date(), estadotutor=2, archivo__isnull=False, estadorevision=1, evidencia__id__in=[9]).order_by('inscripcionpracticas__inscripcion'):
#     contador+=1
#     evidencia.evidencia_id=15
#     evidencia.save()
#     print((contador).__str__()+" - "+evidencia.inscripcionpracticas.inscripcion.persona.__str__() + " - "+ evidencia.inscripcionpracticas.inscripcion.carrera.__str__())

# contador = 0
# for evidencia in DetalleEvidenciasPracticasPro.objects.filter(inscripcionpracticas__fechadesde__gte=datetime.strptime('01-06-2018', "%d-%m-%Y").date(), evidencia__id__in=[15]).distinct('inscripcionpracticas__id').order_by( 'inscripcionpracticas__id'):
#     numero = evidencia.inscripcionpracticas.detalleevidenciaspracticaspro_set.filter(evidencia__id__in=[15]).order_by('fechainicio')
#     if numero.count()>1:
#         contador += 1
#         # print((contador).__str__() + " - " + evidencia.inscripcionpracticas.inscripcion.persona.__str__() + " - " + evidencia.inscripcionpracticas.inscripcion.carrera.__str__() +'  -archivo 1-  '+('SI'if numero[0].archivo else 'NO')+' FECHA '+numero[0].fechainicio.__str__()+'-'+numero[0].fechafin.__str__()+'  -archivo 2-  '+('SI'if numero[1].archivo else 'NO')+' FECHA '+numero[1].fechainicio.__str__()+'-'+numero[1].fechafin.__str__())
#         evidencia1 = numero[0]
#         evidencia1.delete()
#         print((contador).__str__() + " - " + evidencia.inscripcionpracticas.inscripcion.persona.__str__())

# periodoevidencia1 = PeriodoEvidenciaPracticaProfesionales.objects.all()[0]
# periodoevidencia2 = PeriodoEvidenciaPracticaProfesionales.objects.all()[1]
# periodoevidencia3 = PeriodoEvidenciaPracticaProfesionales.objects.all()[2]
# for detalle in DetalleEvidenciasPracticasPro.objects.filter(evidencia__id=20, evidencia__periodoevidencia=periodoevidencia3, archivo__isnull=False):
#     print(detalle.inscripcionpracticas.inscripcion)
# print(DetalleEvidenciasPracticasPro.objects.filter(evidencia__id=20, evidencia__periodoevidencia=periodoevidencia3, archivo__isnull=False).count())
# ASIGNAR PERIODO DE EVIDENCIA A PRACTICAS PRE PROFESIONALES
# contador = 0
# for practica in PracticasPreprofesionalesInscripcion.objects.filter(status=True).distinct():
#     if practica.fechadesde:
#         nevidencia = EvidenciaPracticasProfesionales.objects.values('id').filter(status=True, fechainicio__lte=practica.fechadesde, fechafin__gte=practica.fechadesde).count()
#         if nevidencia == 7:
#             practica.periodoevidencia = periodoevidencia1
#             practica.save()
#             contador+=1
#             print(contador.__str__() + "    -7-     " + practica.fechadesde.__str__()+"     "+ practica.inscripcion.__str__())
#         elif nevidencia == 8:
#             practica.periodoevidencia = periodoevidencia2
#             practica.save()
#             contador += 1
#             print(contador.__str__() + "    -8-     " + practica.fechadesde.__str__() + "     " + practica.inscripcion.__str__())
#         elif nevidencia == 10:
#             practica.periodoevidencia = periodoevidencia3
#             practica.save()
#             contador += 1
#             print(contador.__str__() + "    -10-     " + practica.fechadesde.__str__() + "     " + practica.inscripcion.__str__())

# # CAMBIO DE EVIDENCIA A LAS NUEVA
# for practica in PracticasPreprofesionalesInscripcion.objects.filter(status=True).distinct():
#     if practica.fechadesde:
#         nevidencia = EvidenciaPracticasProfesionales.objects.values('id').filter(status=True, fechainicio__lte=practica.fechadesde, fechafin__gte=practica.fechadesde).count()
#         if nevidencia == 10:
#             evidenciacartavinculacion1 = practica.detalleevidenciaspracticaspro_set.filter(evidencia__id=7)
#             if evidenciacartavinculacion1.exists():
#                 if evidenciacartavinculacion1.count()>1:
#                     print("Tiene mas detalles antes de cambiar de la 7 a la 22     -8-     " + practica.fechadesde.__str__() + "     " + practica.inscripcion.__str__())
#                 for detalle in evidenciacartavinculacion1:
#                     detalle.evidencia_id=22
#                     detalle.save()
#                     contador += 1
#                     print(contador.__str__() + "     "+"Cambiado de la 7 a la 22     -8-     " + practica.fechadesde.__str__() + "     " + practica.inscripcion.__str__())
#
#             evidenciacartavinculacion2 = practica.detalleevidenciaspracticaspro_set.filter(evidencia__id=8)
#             if evidenciacartavinculacion2.exists():
#                 if evidenciacartavinculacion2.count() > 1:
#                     print("Tiene mas detalles antes de cambiar de la 8 a la 23     -8-     " + practica.fechadesde.__str__() + "     " + practica.inscripcion.__str__())
#                 for detalle in evidenciacartavinculacion2:
#                     detalle.evidencia_id = 23
#                     detalle.save()
#                     contador += 1
#                     print(contador.__str__() + "     "+"Cambiado de la 8 a la 23     -8-     " + practica.fechadesde.__str__() + "     " + practica.inscripcion.__str__())
#
#         elif nevidencia == 13:
#             evidenciacartavinculacion3 = practica.detalleevidenciaspracticaspro_set.filter(evidencia__id=14)
#             if evidenciacartavinculacion3.exists():
#                 if evidenciacartavinculacion3.count() > 1:
#                     print("Tiene mas detalles antes de cambiar de la 14 a la 26     -10-     " + practica.fechadesde.__str__() + "     " + practica.inscripcion.__str__())
#                 for detalle in evidenciacartavinculacion3:
#                     detalle.evidencia_id = 26
#                     detalle.save()
#                     contador += 1
#                     print(contador.__str__() + "     "+"Cambiado de la 14 a la 26     -10-     " + practica.fechadesde.__str__() + "     " + practica.inscripcion.__str__())
#
#             evidenciacartavinculacion4 = practica.detalleevidenciaspracticaspro_set.filter(evidencia__id=7)
#             if evidenciacartavinculacion4.exists():
#                 if evidenciacartavinculacion4.count() > 1:
#                     print("Tiene mas detalles antes de cambiar de la 7 a la 24     -10-     " + practica.fechadesde.__str__() + "     " + practica.inscripcion.__str__())
#                 for detalle in evidenciacartavinculacion4:
#                     detalle.evidencia_id = 24
#                     detalle.save()
#                     contador += 1
#                     print(contador.__str__() + "     "+"Cambiado de la 7 a la 24     -10-     " + practica.fechadesde.__str__() + "     " + practica.inscripcion.__str__())
#
#             evidenciacartavinculacion5 = practica.detalleevidenciaspracticaspro_set.filter(evidencia__id=8)
#             if evidenciacartavinculacion5.exists():
#                 if evidenciacartavinculacion5.count() > 1:
#                     print("Tiene mas detalles antes de cambiar de la 8 a la 24     -10-     " + practica.fechadesde.__str__() + "     " + practica.inscripcion.__str__())
#                 for detalle in evidenciacartavinculacion5:
#                     detalle.evidencia_id = 25
#                     detalle.save()
#                     contador += 1
#                     print(contador.__str__() + "     "+"Cambiado de la 8 a la 25     -10-     " + practica.fechadesde.__str__() + "     " + practica.inscripcion.__str__())

#
# contador1=0
# for practica in PracticasPreprofesionalesInscripcion.objects.filter(periodoevidencia__isnull=True, fechadesde__isnull=False):
#     nevidencia = EvidenciaPracticasProfesionales.objects.values('id').filter(status=True, fechainicio__lte=practica.fechadesde, fechafin__gte=practica.fechadesde).count()
#     contador1+=1
#     print(contador1.__str__() + "     " +nevidencia.__str__()+ "     " + practica.fechadesde.__str__() + "     " + practica.inscripcion.__str__())




# print(PracticasPreprofesionalesInscripcion.objects.filter(status=True,fechahasta__gte=datetime.strptime('30-09-2017', "%d-%m-%Y").date(),
# print(PracticasPreprofesionalesInscripcion.objects.filter(status=True, fechadesde__isnull=True).distinct().count())
        # # ACTUALIZAR IDENTIFICACION DE LA MATERIA
# for materia in Materia.objects.filter(nivel__periodo__id__in=[76,77]).order_by('fecha_creacion'):
#     print(materia.fecha_creacion.__str__()+" -  " +materia.__str__())
# # Matricula.objects.filter(inscripcion__carrera__id=22, nivelmalla__id=3).update(cerrada = True)
# print 'listo'

# # abrir la materia
# for materia in Materia.objects.filter(nivel__periodo__id__in=[11], asignaturamalla__malla__carrera__id=3 , cerrado=True).order_by('-id'):
#     materia.cerrado = False
#     materia.save()
#     for asig in materia.asignados_a_esta_materia():
#         asig.cerrado = True
#         asig.save()
#     print materia.id
# print 'listo'

# # pac general
# workbook = xlrd.open_workbook("pac.xlsx")
# sheet = workbook.sheet_by_index(0)
# linea = 1
# periodopac = PeriodoPac.objects.filter(anio=2017)[0]
# periodopoa = PeriodoPoa.objects.filter(anio=2017)[0]
# PacGeneral.objects.all().delete()
# ObjetivosPac.objects.all().delete()
# ProductosPac.objects.all().delete()
# for rowx in range(sheet.nrows):
#     if linea>1:
#         cols = sheet.row_values(rowx)
#         print cols[2].strip()
#         departamento = Departamento.objects.get(pk=int(cols[0].strip()))
#         if not ObjetivosPac.objects.filter(departamento=departamento, descripcion=cols[1].strip().upper(), periodopoa=periodopoa).exists():
#             objetivospac = ObjetivosPac(departamento=departamento,
#                                         descripcion=cols[1].strip().upper(),
#                                         periodopoa=periodopoa)
#             objetivospac.save()
#         else:
#             objetivospac = ObjetivosPac.objects.filter(departamento=departamento, descripcion=cols[1].strip().upper(), periodopoa=periodopoa)[0]
#
#         if not ProductosPac.objects.filter(descripcion=cols[2].strip().upper()).exists():
#             productopac = ProductosPac(descripcion=cols[2].strip().upper())
#             productopac.save()
#         else:
#             productopac = ProductosPac.objects.filter(descripcion=cols[2].strip().upper())[0]
#
#         pacgeneral = PacGeneral(periodo = periodopac,
#                                 objetivospac = objetivospac,
#                                 productospac = productopac,
#                                 total = Decimal(cols[3]).quantize(Decimal('.01')),
#                                 valorejecutado = 0)
#         pacgeneral.save()
#
#         for i in range(12):
#             if cols[4+i] == '':
#                 valor = 0
#             else:
#                 valor = Decimal(cols[4+i]).quantize(Decimal('.01'))
#             pacdetallegeneral = PacDetalladoGeneral(pacgeneral = pacgeneral,
#                                                     mes = i+1,
#                                                     valor = valor)
#             pacdetallegeneral.save()
#
#
#
#     linea += 1
# # print 'listo'
# egresado = Egresado.objects.values_list('inscripcion__id', flat=True)
# admision = Inscripcion.objects.values_list('id', flat=True).filter(carrera__in=CARRERAS_ADMISION)
#
# for inscripcion in Inscripcion.objects.filter(status=True, matricula__nivel__periodo__id=11).order_by('-id').exclude(id__in=egresado).exclude(id__in=admision):
#     inscripcion.actualiza_estado_matricula()
#     print inscripcion.id
# print 'listo'

# # alumnos al primer semestre que ya vienen de admision
# workbook = xlrd.open_workbook("inscripcion_2017_1.xlsx")
# sheet = workbook.sheet_by_index(0)
# linea = 1
# modalidad = Modalidad.objects.get(pk=1)
# sede = Sede.objects.get(pk=1)
# for rowx in range(sheet.nrows):
#     if linea>1:
#         cols = sheet.row_values(rowx)
#         cedula = cols[0].strip().upper()
#         if not Persona.objects.filter(cedula=cedula).exists():
#             print ('cedula no existe %s' % cedula)
#         else:
#             persona = Persona.objects.filter(cedula=cedula)[0]
#             sesion = Sesion.objects.get(pk=int(cols[3]))
#             carrera = Carrera.objects.get(pk=int(cols[2]))
#             colegio = persona.inscripcion_set.all()[0].colegio
#             raza_id = 6
#             if not Inscripcion.objects.filter(persona=persona,carrera=carrera).exists():
#                 inscripcion = Inscripcion(persona=persona,
#                                           fecha=datetime.now().date(),
#                                           carrera=carrera,
#                                           modalidad=modalidad,
#                                           sesion=sesion,
#                                           sede=sede,
#                                           colegio=colegio)
#                 inscripcion.save()
#                 persona.crear_perfil(inscripcion=inscripcion)
#                 documentos = DocumentosDeInscripcion(inscripcion=inscripcion,
#                                                      titulo=False,
#                                                      acta=False,
#                                                      cedula=False,
#                                                      votacion=False,
#                                                      actaconv=False,
#                                                      partida_nac=False,
#                                                      pre=False,
#                                                      observaciones_pre='',
#                                                      fotos=False)
#                 documentos.save()
#                 preguntasinscripcion = inscripcion.preguntas_inscripcion()
#                 perfil_inscripcion = inscripcion.persona.mi_perfil()
#
#                 perfil_inscripcion.raza_id =  raza_id
#                 perfil_inscripcion.save()
#                 inscripciontesdrive = InscripcionTesDrive(inscripcion=inscripcion,
#                                                           licencia=False,
#                                                           record=False,
#                                                           certificado_tipo_sangre=False,
#                                                           prueba_psicosensometrica=False,
#                                                           certificado_estudios=False)
#                 inscripciontesdrive.save()
#                 # inscripcion.mi_malla()
#                 inscripcion.malla_inscripcion()
#                 inscripcion.actualizar_nivel()
#                 if USA_TIPOS_INSCRIPCIONES:
#                     inscripciontipoinscripcion = InscripcionTipoInscripcion(inscripcion=inscripcion,
#                                                                             tipoinscripcion_id=TIPO_INSCRIPCION_INICIAL)
#                     inscripciontipoinscripcion.save()
#                 # persona.creacion_persona(request.session['nombresistema'])
#             else:
#                 inscripcion = Inscripcion.objects.filter(persona=persona,carrera=carrera)[0]
#                 perfil_inscripcion = inscripcion.persona.mi_perfil()
#                 perfil_inscripcion.raza_id = raza_id
#                 perfil_inscripcion.save()
#     linea += 1
#     print (linea)



# # alumnos con titulos superior
# workbook = xlrd.open_workbook("reporte_titulos_unemi.xlsx")
# sheet = workbook.sheet_by_index(0)
# linea = 1
# for rowx in range(sheet.nrows):
#     if linea>1:
#         cols = sheet.row_values(rowx)
#         cedula = cols[0].strip().upper()
#         persona = Persona.objects.get(cedula=cedula)
#         if persona:
#             fecharegistro = None
#             if cols[2].strip() != '':
#                 fecharegistro = convertirfecha2(cols[2].strip())
#
#             fechainicio = None
#             if cols[3].strip() != '':
#                 fechainicio = convertirfecha2(cols[3].strip())
#
#             fecharegresado = None
#             if cols[4].strip() != '':
#                 fecharegresado = convertirfecha2(cols[4].strip())
#
#             fechaacta = None
#             if cols[5].strip() != '':
#                 fechaacta = convertirfecha2(cols[5].strip())
#
#             universidad = None
#             if InstitucionEducacionSuperior.objects.filter(codigo=cols[6].strip()).exists():
#                 universidad = InstitucionEducacionSuperior.objects.filter(codigo=cols[6].strip())[0]
#             else:
#                 universidad = InstitucionEducacionSuperior(nombre = cols[7].strip(),
#                                                            codigo = cols[6].strip())
#                 universidad.save()
#
#             if not PersonaTituloUniversidad.objects.filter(persona = persona, codigoregistro = cols[1].strip(),universidad = universidad).exists():
#                 personatitulouniversidad = PersonaTituloUniversidad(persona = persona,
#                                                                     codigoregistro = cols[1].strip(),
#                                                                     fecharegistro = fecharegistro,
#                                                                     fechainicio = fechainicio,
#                                                                     fecharegresado = fecharegresado,
#                                                                     fechaacta = fechaacta,
#                                                                     universidad = universidad,
#                                                                     tipouniversidad = int(cols[8]),
#                                                                     nombrecarrera = cols[9].strip(),
#                                                                     tiponivel = int(cols[10]))
#                 personatitulouniversidad.save()
#         else:
#             print 'cedula no existe %s' % cedula
#
#     linea += 1



# # matriculacion de alumnos
# workbook = xlrd.open_workbook("inicialprimersemestre.xlsx")
# sheet = workbook.sheet_by_index(0)
# linea = 1
# periodo = Periodo.objects.get(pk=14)
# modalidad = Modalidad.objects.get(pk=1)
# for rowx in range(sheet.nrows):
#     if linea>1:
#         cols = sheet.row_values(rowx)
#         cedula = cols[0].strip().upper()
#         bandera = True
#         if Persona.objects.filter(cedula=cedula).exists():
#             persona = Persona.objects.filter(cedula=cols[0])[0]
#             sesion = Sesion.objects.get(pk=int(cols[2]))
#             carrera = Carrera.objects.get(pk=int(cols[1]))
#             sede = Sede.objects.get(pk=1)
#             # cordinacion_alias = str(carrera.coordinacion_set.all()[0].alias)
#             cordinacion_alias = 'FESAD'
#             colegio = persona.inscripcion_set.all()[0].colegio
#             raza_id = 6
#             if not Inscripcion.objects.filter(persona=persona,carrera=carrera).exists():
#                 inscripcion = Inscripcion(persona=persona,
#                                           fecha=datetime.now().date(),
#                                           carrera=carrera,
#                                           modalidad=modalidad,
#                                           sesion=sesion,
#                                           sede=sede,
#                                           colegio=colegio)
#                 inscripcion.save()
#                 persona.crear_perfil(inscripcion=inscripcion)
#                 documentos = DocumentosDeInscripcion(inscripcion=inscripcion,
#                                                      titulo=False,
#                                                      acta=False,
#                                                      cedula=False,
#                                                      votacion=False,
#                                                      actaconv=False,
#                                                      partida_nac=False,
#                                                      pre=False,
#                                                      observaciones_pre='',
#                                                      fotos=False)
#                 documentos.save()
#                 preguntasinscripcion = inscripcion.preguntas_inscripcion()
#                 perfil_inscripcion = inscripcion.persona.mi_perfil()
#
#                 perfil_inscripcion.raza_id =  raza_id
#                 perfil_inscripcion.save()
#                 inscripciontesdrive = InscripcionTesDrive(inscripcion=inscripcion,
#                                                           licencia=False,
#                                                           record=False,
#                                                           certificado_tipo_sangre=False,
#                                                           prueba_psicosensometrica=False,
#                                                           certificado_estudios=False)
#                 inscripciontesdrive.save()
#                 # inscripcion.mi_malla()
#                 inscripcion.malla_inscripcion()
#                 inscripcion.actualizar_nivel()
#                 if USA_TIPOS_INSCRIPCIONES:
#                     inscripciontipoinscripcion = InscripcionTipoInscripcion(inscripcion=inscripcion,
#                                                                             tipoinscripcion_id=TIPO_INSCRIPCION_INICIAL)
#                     inscripciontipoinscripcion.save()
#                 # persona.creacion_persona(request.session['nombresistema'])
#             else:
#                 inscripcion = Inscripcion.objects.filter(persona=persona,carrera=carrera)[0]
#                 perfil_inscripcion = inscripcion.persona.mi_perfil()
#                 perfil_inscripcion.raza_id = raza_id
#                 perfil_inscripcion.save()
#
#
#
#             inscripcion = Inscripcion.objects.filter(persona=persona,carrera=carrera)[0]
#             inscripcion.sesion = sesion
#             inscripcion.save()
#
#         # considerar si que existe en titulo persona
#         if not persona.tiene_otro_titulo():
#             nivel = Nivel.objects.get(periodo=periodo, sesion=sesion, paralelo__icontains=cordinacion_alias)
#             # matricula
#             if not inscripcion.matricula_periodo(periodo):
#                 matricula = Matricula(inscripcion=inscripcion,
#                                       nivel=nivel,
#                                       pago=False,
#                                       iece=False,
#                                       becado=False,
#                                       porcientobeca=0,
#                                       fecha=datetime.now().date(),
#                                       hora=datetime.now().time(),
#                                       fechatope=fechatope(datetime.now().date()))
#                 matricula.save()
#             else:
#                 matricula = Matricula.objects.get(inscripcion=inscripcion, nivel=nivel)
#
#             asignatura_malla=RecordAcademico.objects.values_list('asignatura__id',flat=True).filter(asignaturamalla__nivelmalla__id__lt=int(cols[4]), aprobada=True)
#
#             for materia in Materia.objects.filter(nivel__periodo=periodo, paralelo=str(cols[3].strip()), asignaturamalla__malla__carrera=carrera, nivel__sesion=sesion, asignaturamalla__nivelmalla__id=int(cols[4])):
#                 if not MateriaAsignada.objects.filter(matricula=matricula,materia=materia).exists():
#                     matriculas = matricula.inscripcion.historicorecordacademico_set.filter(asignatura=materia.asignatura, fecha__lt=materia.nivel.fin).count() + 1
#                     materiaasignada = MateriaAsignada(matricula=matricula,
#                                                       materia=materia,
#                                                       notafinal=0,
#                                                       asistenciafinal=0,
#                                                       cerrado=False,
#                                                       matriculas=matriculas,
#                                                       observaciones='',
#                                                       estado_id=NOTA_ESTADO_EN_CURSO)
#                     materiaasignada.save()
#                     materiaasignada.asistencias()
#                     materiaasignada.evaluacion()
#                     materiaasignada.mis_planificaciones()
#                     materiaasignada.save()
#             matricula.actualizar_horas_creditos()
#             matricula.actualiza_matricula()
#     linea += 1
#     print (linea)


# sacar estudiando pasados limpios
# periodo = Periodo.objects.get(pk=9)
# linea = 0
# for inscripcion in Inscripcion.objects.filter(matricula__nivel__periodo=periodo, coordinacion_id__in=[1]).order_by('coordinacion', 'carrera'):
#     linea += 1
#     malla = inscripcion.carrera.malla()
#     nivelmalla_matricula = inscripcion.matricula_periodo(periodo).nivelmalla_id
#     asignaturas_atras = AsignaturaMalla.objects.filter(nivelmalla__id__lte=nivelmalla_matricula, malla=malla)
#     asignaturas_delante = AsignaturaMalla.objects.filter(nivelmalla__id__gt=nivelmalla_matricula, malla=malla)
#     bandera = 0
#     for asignatura_atras in asignaturas_atras:
#         if bandera == 0:
#             if RecordAcademico.objects.filter(inscripcion=inscripcion, asignatura=asignatura_atras.asignatura).exists():
#                 record = RecordAcademico.objects.filter(inscripcion=inscripcion, asignatura=asignatura_atras.asignatura)[0]
#                 if record.aprobada == False:
#                     bandera = 1
#             else:
#                 bandera = 1
#     if bandera == 0:
#         for asignatura_delante in asignaturas_delante:
#             if RecordAcademico.objects.filter(inscripcion=inscripcion, asignatura=asignatura_delante.asignatura).exists():
#                 bandera = 1
#
#     if bandera == 0:
#         data = [[inscripcion.persona.cedula, malla.carrera_id, inscripcion.sesion_id, "", nivelmalla_matricula]]
#         dataWriter.writerows(data)
#         print inscripcion
#     else:
#         print "%s - %s" % (linea, inscripcion.persona.cedula)
# print 'listo'

# for matricula in Matricula.objects.filter(nivel__periodo_id=11, estado_matricula=1).order_by('-id'):
#     matricula.actualiza_matricula()
#     print matricula.id
# print 'listo'


# for traspasos in TraspasoActivo.objects.filter(tipo=2, status=True).order_by('estado', '-numero'):
#     traspasos.totalbienes = traspasos.cantidad_seleccionados()
#     traspasos.save(usuario=traspasos.usuario_modificacion)
#     print traspasos.id



# # sacar estudiando pasados en todo para ver si estan en egresados
# linea = 0
# for inscripcion in Inscripcion.objects.filter(status=True).order_by('coordinacion', 'carrera'):
#     linea += 1
#     malla = inscripcion.carrera.malla()
#     asignaturas = AsignaturaMalla.objects.filter(malla=malla)
#     if asignaturas.count() > 0:
#         bandera = 0
#         for asignatura in asignaturas:
#             if bandera == 0:
#                 if not RecordAcademico.objects.filter(inscripcion=inscripcion, asignatura=asignatura.asignatura, aprobada = True, status=True).exists():
#                     bandera = 1
#         if bandera == 0:
#             if Egresado.objects.filter(inscripcion=inscripcion, status=True).exists():
#                 bandera = 1
#         if bandera == 0:
#             data = [[inscripcion.persona.cedula , remover_caracteres_especiales(inscripcion.persona.nombre_completo_inverso()), malla.carrera, inscripcion.persona.email, inscripcion.persona.emailinst , inscripcion.persona.telefono_conv, inscripcion.persona.telefono, remover_caracteres_especiales(inscripcion.persona.direccion_completa()), inscripcion.persona.nacionalidad, inscripcion.sesion]]
#             dataWriter.writerows(data)
#             print "%s - %s - %s" % (inscripcion.persona.cedula, malla.carrera, inscripcion.sesion)
#     print linea
# print 'listo'


# # actualizacion de malla
# for asignaturamalla in AsignaturaMalla.objects.filter(malla__id=32):
#     RecordAcademico.objects.filter(asignatura=asignaturamalla.asignatura).update(creditos=asignaturamalla.creditos,horas=asignaturamalla.horas)
#     HistoricoRecordAcademico.objects.filter(asignatura=asignaturamalla.asignatura).update(creditos=asignaturamalla.creditos,horas=asignaturamalla.horas)
#     # print asignaturamalla
#
# # actualiza credito
# for inscripcion in Inscripcion.objects.filter(carrera__id=1):
#     inscripcion.actualizar_creditos()
#     inscripcion.actualizar_niveles_records()
#     inscripcion.actualizar_nivel()
#     print inscripcion

# # actualizacion kardex
# for detalle in DetalleIngresoProducto.objects.all():
#     ingreso = detalle.ingreso_producto()
#     if detalle.kardexinventario_set.exists():
#         kardex = detalle.kardex()
#         if ingreso.anulado:
#             kardex.anulado = True
#             kardex.save()
#             print kardex.id
#
# for detalle in DetalleSalidaProducto.objects.all():
#     salida = detalle.salida_producto()
#     if detalle.kardexinventario_set.exists():
#         kardex = detalle.kardex_inventario()
#         if salida.anulado:
#             kardex.anulado = True
#             kardex.save()
#             print kardex.id
#
#

# # inscripcion
# workbook = xlrd.open_workbook("estudiantes.xlsx")
# sheet = workbook.sheet_by_index(0)
# linea = 1
# for rowx in range(sheet.nrows):
#     print linea
#     if linea>1:
#         cols = sheet.row_values(rowx)
#         carrera_id = str(int(cols[0])).upper()
#         cedula = cols[1].strip().upper()
#         fechainicioprimernivel = None
#         if cols[2] != '':
#             fechainicioprimernivel = convertirfecha2(cols[2])
#
#         fechainicioconvalidacion = None
#         if cols[3] != '':
#             fechainicioconvalidacion = convertirfecha2(cols[3])
#
#         fechagraduado = None
#         if cols[4] != '':
#             fechagraduado = convertirfecha2(cols[4])
#
#         if Persona.objects.filter(Q(cedula=cedula) | Q(pasaporte=cedula)).exists():
#             persona = Persona.objects.filter(Q(cedula=cedula) | Q(pasaporte=cedula))[0]
#             if Inscripcion.objects.filter(persona=persona, carrera__id=carrera_id).exists():
#                 inscripcion = Inscripcion.objects.filter(persona=persona, carrera__id=carrera_id)[0]
#                 inscripcion.fechainicioprimernivel = fechainicioprimernivel
#                 inscripcion.fechainicioconvalidacion = fechainicioconvalidacion
#                 inscripcion.save()
#                 if fechagraduado:
#                     if inscripcion.graduado():
#                         graduado = inscripcion.graduado_set.all()[0]
#                         graduado.fechagraduado = fechagraduado
#                         graduado.save()
#                     else:
#                         print 'no existe graduado %s' % cedula
#             else:
#                 print 'no existe inscripcion %s' % cedula
#         else:
#             print 'no existe cedula %s' % cedula
#
#
#     linea += 1
#


# for matriculas in Matricula.objects.filter(nivel__periodo__id=11):
#     print matriculas.id
#     matriculas.actualiza_matricula()
# print 'listo'


# # cagada locke yepez
# cabeceras_si = [1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,22,23,24,25,26,27,29,30,38,39,40]
# cabeceras_no = [41,42,43,44,45,46,47,48,49,50,51,52,53,54,55,56,57,58,59,60,61,62,63,64,65,66,67,68,69,70,71,72,73,74,75,76,77,78,79,80,81,82,83,84,85,86,87,88,89,90,91,92,93,94,95,96,97,98,99,100,101,102,103,104,105,106,107,108,109,110,111,112,113,114,115,116,117,118,119,120,121,122,123,124,125,126,127,128,129,130,131,132,133,134,135,136,137,138,139,140,141,142,143,144,145,146,147,148,149,150,151,152,153,154,155,156,157,158,159,160,161,162,163,164,165,166,167,168,169,170,171,172,173,174,175,176,177,178,179,180,181,182,183,184,185,186,187,188,189,190,191,192,193,194,195,196,197,198,199,200,201,202,203,204,205,206,207,208,209,210,211,212,213,214,215,216]
# contador_respuestas=cabeceras_si.__len__()
# i = 0
# for cabecera_no in cabeceras_no:
#     if i >= contador_respuestas:
#         i = 0
#     for r in SagResultadoEncuestaDetalle.objects.filter(sagresultadoencuesta__id=int(cabeceras_si[i])):
#         sagresultadoencuestadetalle = SagResultadoEncuestaDetalle(sagresultadoencuesta_id = int(cabecera_no),
#                                                                   preguntaencuesta = r.preguntaencuesta,
#                                                                   valor = r.valor,
#                                                                   numero = r.numero)
#         sagresultadoencuestadetalle.save()
#     print "Cabecera %s, se la va poner la respuesta %s " % (cabecera_no, cabeceras_si[i])
#     i +=1

# # afinidad
# ProfesorMateria.objects.all().update(afinidad=False)
# for periodo in Periodo.objects.filter(status=True).exclude(tipo__id=1).order_by('id'):
#     print periodo
#     for profesormateria in ProfesorMateria.objects.filter(materia__nivel__periodo=periodo).order_by('profesor'):
#         asignaturamallatituloafin = AsignaturaMallaTituloAFin.objects.values_list('titulo__id',flat=True).filter(status=True, asignaturamalla=profesormateria.materia.asignaturamalla)
#         for titulo in profesormateria.profesor.persona.titulacion_set.filter(status=True, titulo__id__in=asignaturamallatituloafin).exclude(fechaobtencion__gt=periodo.inicio):
#             print "%s - %s - %s - %s" % (profesormateria.profesor.persona, titulo, profesormateria.materia.asignatura, profesormateria.materia.asignaturamalla.malla.carrera)
#             ProfesorMateria.objects.filter(id=profesormateria.id).update(afinidad=True)


# for practicas in PracticasPreprofesionalesInscripcion.objects.filter(status=True):
#     if DetalleEvidenciasPracticasPro.objects.filter(inscripcionpracticas=practicas, status=True).exists():
#         practicas.culminada=False
# print 'listo'

# arregal malla de computacion
# workbook = xlrd.open_workbook("computacioncredito.xlsx")
# sheet = workbook.sheet_by_index(0)
# linea = 1
# for rowx in range(sheet.nrows):
#     print linea
#     if linea>1:
#         cols = sheet.row_values(rowx)
#         if AsignaturaMalla.objects.filter(malla__id = 32, asignatura__nombre=cols[1].strip(), status=True).exists():
#             AsignaturaMalla.objects.filter(malla__id = 32, asignatura__nombre=cols[1].strip(), status=True).update(creditos=float(cols[3]))
#         else:
#             print cols[1].strip()
#     linea += 1

# asignatura1 = Asignatura.objects.filter(id=786)[0]
# RecordAcademico.objects.filter(inscripcion__carrera__id=7, status=True, asignatura__id=1679).update(asignatura=asignatura1)
# HistoricoRecordAcademico.objects.filter(inscripcion__carrera__id=7, status=True, asignatura__id__in=[1679]).update(asignatura=asignatura1)


# for inscripcion in Inscripcion.objects.filter(matricula__nivel__periodo__id=11):
#     materiasasignada = MateriaAsignada.objects.values_list('id', flat=True).filter(matricula__inscripcion=inscripcion, matricula__nivel__periodo__id=11)
#     for materiaasignadaplanificacion in MateriaAsignadaPlanificacion.objects.filter(materiaasignada__matricula__inscripcion=inscripcion, planificacion__materia__nivel__periodo__id=11, status=True, materiaasignada__matricula__nivel__periodo__id=11).exclude(id__in=materiasasignada):
#         print '%s - %s' % (inscripcion, materiaasignadaplanificacion)


# # deudas devolucion
# workbook = xlrd.open_workbook("NotificacionesEstudiantes.xlsx")
# sheet = workbook.sheet_by_index(0)
# n = 1
# codigo = ''
#
# tiporubro = None
# if TipoOtroRubro.objects.filter(pk=2951).exists():
#     tiporubro = TipoOtroRubro.objects.filter(pk=2951)[0]
#
# for rowx in range(sheet.nrows):
#     print n
#     if n > 1:
#         row = sheet.row_values(rowx)
#         persona = None
#         if Persona.objects.filter(cedula=row[0].strip()).exists():
#             persona = Persona.objects.filter(cedula=row[0].strip())[0]
#
#         rubro = Rubro(tipo=tiporubro,
#                       persona = persona,
#                       relacionados = None,
#                       matricula = None,
#                       # contratorecaudacion = None,
#                       nombre = tiporubro.nombre,
#                       cuota = 1,
#                       fecha = convertirfecha2(row[2].strip()),
#                       fechavence = convertirfecha2(row[3].strip()),
#                       valor = Decimal(row[1]).quantize(Decimal('.01')),
#                       iva_id = 1,
#                       valoriva = 0,
#                       valortotal = Decimal(row[1]).quantize(Decimal('.01')),
#                       saldo = Decimal(row[1]).quantize(Decimal('.01')),
#                       cancelado = False)
#         rubro.save()
#
#     n += 1
# print 'listo'

# #  Revision de Titulos
# for titulo in Titulo.objects.filter(status=True, nivel__id=4).order_by('id'):
#     if not Titulacion.objects.filter(titulo=titulo, status=True).exists():
#         print '%s - %s' % (titulo.id, titulo)
# periodo = Periodo.objects.get(pk=11)
# sede = Sede.objects.get(pk=1)
# for c in ComplexivoTematica.objects.filter(status=True, director__isnull=True, periodo__isnull=False):
#     print c.id
#     print c.director
#     print c.carrera
#     print c.carrera.coordinador_carrera(periodo, sede)[0].persona.profesor()
#     c.director=c.carrera.coordinador_carrera(periodo, sede)[0].persona.profesor()
#     c.save()

# arregla kardex ay que poner al comienzo un id menor para que funcione
# producto = Producto.objects.get(pk=13)
# idanterior = 16240
# idnormal = 16270
# arregloid = KardexInventario.objects.values_list('id', flat=True).filter(producto=producto, id__gte=idanterior, status=True).order_by('id')
# i = 0
# for k in KardexInventario.objects.filter(producto=producto, id__gte=idnormal, status=True).order_by('id'):
#     print (k.id)
#     saldoanterior = KardexInventario.objects.get(pk=arregloid[i]).saldofinalvalor
#     cantidadanterior = KardexInventario.objects.get(pk=arregloid[i]).saldofinalcantidad
#     print (saldoanterior)
#     k.saldoinicialvalor = saldoanterior
#     k.saldoinicialcantidad = cantidadanterior
#     if k.tipomovimiento==2:
#         k.saldofinalvalor = saldoanterior - k.valor
#         k.saldofinalcantidad = cantidadanterior - k.cantidad
#     else:
#         k.saldofinalvalor = saldoanterior + k.valor
#         k.saldofinalcantidad = cantidadanterior + k.cantidad
#     k.save()
#     i += 1
# kardex = KardexInventario.objects.filter(producto=producto, status=True).order_by('-id')[0]
# valorproducto = kardex.saldofinalvalor
# cantidadproducto = kardex.saldofinalcantidad
# costoproducto = kardex.costo
# InventarioReal.objects.filter(producto=producto, status=True).update(valor=valorproducto, cantidad=cantidadproducto, costo=costoproducto)



# # sacar estudiando pasados limpios
# periodo = Periodo.objects.get(pk=11)
# linea = 0
# for inscripcion in Inscripcion.objects.filter(matricula__nivel__periodo=periodo, matricula__nivelmalla__gte=6).order_by('coordinacion', 'carrera'):
#     linea += 1
#     malla = inscripcion.carrera.malla()
#     nivelmalla_matricula = inscripcion.matricula_periodo(periodo).nivelmalla_id-1
#     asignaturas_atras = AsignaturaMalla.objects.filter(nivelmalla__id__lte=nivelmalla_matricula, malla=malla)
#     bandera = 0
#     for asignatura_atras in asignaturas_atras:
#         if bandera == 0:
#             if RecordAcademico.objects.filter(inscripcion=inscripcion, asignatura=asignatura_atras.asignatura).exists():
#                 record = RecordAcademico.objects.filter(inscripcion=inscripcion, asignatura=asignatura_atras.asignatura)[0]
#                 if record.aprobada == False:
#                     bandera = 1
#                 if record.aprobada==True and record.nota<80:
#                     bandera = 1
#             else:
#                 bandera = 1
#     if bandera == 0:
#         print "%s - %s" % (linea, inscripcion.persona.cedula)
# print 'listo'

# # arreglar lo de tipo matricula
# for matricula in Matricula.objects.filter(nivel__periodo__id=14):
#     cursor = connection.cursor()
#     sql = "select am.nivelmalla_id, count(am.nivelmalla_id) as cantidad_materias_seleccionadas " \
#           " from sga_materiaasignada ma, sga_materia m, sga_asignaturamalla am where ma.status=true and ma.matricula_id=" + str(matricula.id) + " and m.status=true and m.id=ma.materia_id and am.status=true and am.id=m.asignaturamalla_id GROUP by am.nivelmalla_id, am.malla_id order by count(am.nivelmalla_id) desc, am.nivelmalla_id desc limit 1;"
#     cursor.execute(sql)
#     results = cursor.fetchall()
#     nivel = 0
#     for per in results:
#         nivel = per[0]
#         cantidad_seleccionadas = per[1]
#     cantidad_nivel = 0
#
#     for asignaturamalla in AsignaturaMalla.objects.filter(nivelmalla__id=nivel, status=True, malla=matricula.inscripcion.mi_malla()):
#         if Materia.objects.filter(nivel__periodo=matricula.nivel.periodo, asignaturamalla=asignaturamalla).exists():
#             if matricula.inscripcion.estado_asignatura(asignaturamalla.asignatura) != 1:
#                 cantidad_nivel += 1
#
#     porcentaje_seleccionadas = int(round(Decimal((float(cantidad_nivel) * float(PORCIENTO_PERDIDA_PARCIAL_GRATUIDAD)) / 100).quantize(Decimal('.00')), 0))
#     cobro = 0
#     if (cantidad_seleccionadas < porcentaje_seleccionadas):
#         cobro = 2
#     else:
#         # if self.inscripcion.estado_gratuidad == 2:
#         cobro = 1
#     matricula.grupo_socio_economico(cobro)
#     print (matricula)

# nocoordinaciones = Coordinacion.objects.filter(id__in=[7,9])
# for materia in MateriaAsignada.objects.filter(asistenciafinal__lt=70,status=True, matricula__nivel__periodo__id=11, matriculas__gt=1, sinasistencia=False).exclude(materia__asignaturamalla__malla__carrera__id__in=CARRERAS_ADMISION).order_by('materia__asignaturamalla__malla__carrera'):
#     if materia.matricula.inscripcion.historicorecordacademico_set.filter(completoasistencia=True, aprobada=False, fecha__lt=materia.materia.inicio, asignatura=materia.materia.asignatura).exists():
#         if materia.evaluaciongenerica_set.filter(status=True, detallemodeloevaluativo__nombre__icontains='EX')[0].valor==0:
#             materia.sinasistencia=True
#             materia.save()
#             print materia


# modificar paralelo en materia
# mat=Materia.objects.filter(status=True)
# for m in mat:
#     # paral=m.paralelo.replace("'","")
#     if not Paralelo.objects.select_related().filter(nombre=m.paralelo):
#         paral=Paralelo(nombre=m.paralelo)
#         paral.save()
#         print (paral.nombre)



for matricula in Matricula.objects.filter(status=True,nivel__periodo__id__in=[85]):
    try:
        paralelos = MateriaAsignada.objects.filter(status=True, matricula=matricula,matricula__nivel__periodo__id__in=[85])
        listparalelo = []
        for x in paralelos:
            a = x.materia.paralelo
            listparalelo.append(a)
        cuenta1 = collections.Counter(listparalelo).most_common(1)
        if cuenta1:
            paralelo=cuenta1[0][0]
            if Paralelo.objects.select_related().filter(nombre=paralelo).exists():
                idparaleo=Paralelo.objects.select_related().filter(nombre=paralelo)
                for p in idparaleo:
                    matricula.paralelo_id = p.id
                    matricula.save()
            print (matricula,paralelo)
    except Exception as ex:
        pass

# # CAMBIO DE CARRERA A OTRA COORDINACION
# carreraacambiar=46
# idasignaturasmallas = AsignaturaMalla.objects.values_list('id', flat=True).filter(status=True, malla__carrera__id=carreraacambiar)
# print(idasignaturasmallas)
# for periodo in Periodo.objects.filter(status=True, tipo__id=2).order_by('id'):
#     nivelacambiar = Nivel.objects.filter(status=True, paralelo__icontains='FESAD', periodo=periodo)
#     if nivelacambiar:
#         print(periodo)
#         print("seccion fin de semana %s - %s" % (nivelacambiar[0], nivelacambiar[0].id))
#         for nivel in Nivel.objects.filter(periodo=periodo, status=True, materia__asignaturamalla__id__in=idasignaturasmallas).distinct().order_by('id'):
#             print(nivel)
#             Materia.objects.filter(status=True, nivel=nivel , asignaturamalla__id__in=idasignaturasmallas).update(nivel=nivelacambiar[0])


# eliminar materias del distributivo cuando ya ay matriculados
# materia = Materia.objects.filter(pk=8467)[0]
# print(materia)
# for materiaasignada in MateriaAsignada.objects.filter(status=True, materia=materia):
#     historico = HistoricoRecordAcademico.objects.filter(asignatura=materiaasignada.materia.asignatura, status=True, inscripcion=materiaasignada.matricula.inscripcion)
#     if historico.count()>1:
#         record = RecordAcademico.objects.filter(asignatura=materiaasignada.materia.asignatura, status=True, inscripcion=materiaasignada.matricula.inscripcion)[0]
#         historico = record.historicorecordacademico_set.filter(status=True).order_by('-id')[0]
#         historico.delete()
#         historicoaux = record.historicorecordacademico_set.filter(status=True).order_by('-id')[0]
#         historicoaux.actualizar()
#         print(materiaasignada)
#     else:
#         RecordAcademico.objects.filter(asignatura=materiaasignada.materia.asignatura, status=True, inscripcion=materiaasignada.matricula.inscripcion).delete()
#         materiaasignada.delete()
# materia.delete()


# # conflito de horarios de los alumnos
# periodo = Periodo.objects.filter(pk=14)[0]
# for matricula in Matricula.objects.filter(status=True, nivel__periodo=periodo).order_by('inscripcion__carrera', 'inscripcion'):
#     materias = matricula.materias_periodo()
#     conflicto = conflicto_materias_seleccionadas(materias)
#     if conflicto:
#         print(matricula)
#         print(conflicto)


# # conflito de horarios de los semestre
# nivelesid = [1,2,3,4,5,6,7,8,9]
# periodo = Periodo.objects.filter(pk=14)[0]
# for nivel in Nivel.objects.filter(status=True, periodo=periodo).order_by('id'):
#     materias = matricula.materias_periodo()
#     conflicto = conflicto_materias_seleccionadas(materias)
#     if conflicto:
#         print(matricula)
#         print(conflicto)


# cedulas = ['0940125552','0956228449','0929761120','0920024510','0941880502','0928805530','0958458911','0928474063','0956073332','0941677437','0940153588','1724534639','1803859469','0927146738','0941880692','0930316229']
# for inscripcion in Inscripcion.objects.filter(status=True, matriculacursoescuelacomplementaria__curso__id=101).order_by('id'):
#     record = RecordAcademico.objects.filter(inscripcion=inscripcion, asignatura__id=983)[0]
#     print(inscripcion)
#     record.historicorecordacademico_set.filter(fecha='2017-09-29').delete()
#     historico=record.historicorecordacademico_set.all()[0]
#     historico.actualizar()


# # actualizar estado de la matricula
# for matricula in Matricula.objects.filter(nivel__periodo__id=14):
#     matricula.actualiza_matricula()
#     print(matricula)




# a = 0
# miarchivo = openpyxl.load_workbook('primeros_2012.xlsx')
# lista = miarchivo.get_sheet_by_name('estudiantes')
# totallista = lista.rows
# for filas in totallista[:]:
#     a += 1
#     if a > 1:
#         if Inscripcion.objects.filter(persona__cedula=filas[12].value.strip(),coordinacion__id=filas[4].value,carrera__id=filas[6].value).exists():
#             inscripcion = Inscripcion.objects.get(persona__cedula=filas[12].value.strip(),coordinacion__id=filas[4].value,carrera__id=filas[6].value)
#             inscripcion.fechainicioprimernivel = filas[3].value
#             inscripcion.save()
#             filas[11].value = inscripcion.id
#             print(filas[3].value)
# miarchivo.save("primeros_2012.xlsx")

# import openpyxl
# doc = openpyxl.load_workbook('listaestudiantes_todos.xlsx')
# sheet = doc.get_sheet_by_name('periodos')
# all_rows = sheet.rows
# a = 0
# for s in all_rows[:]:
#     a += 1
#     if a>1:
#         if not Periodo.objects.filter(nombre=s[1].value).exists():
#             periodos = Periodo(nombre=s[1].value,inicio=s[2].value,fin=s[3].value,tipo_id=s[4].value,
#                                inicio_agregacion=datetime.now().date(),limite_agregacion=datetime.now().date(),
#                                limite_retiro=datetime.now().date()
#                                )
#             periodos.save()
#             print('ingresando periodo ' + str(a))
        # else:
        #     periodo = Periodo.objects.get(nombre=s[1].value)
        #     s[0].value = periodo.id
        #     doc.save("listaestudiantes_todos.xlsx")


# /*actualiza archivo*/
# import openpyxl
# a = 0
# miarchivo = openpyxl.load_workbook('listaestudiantes_todos.xlsx')
# lista = miarchivo.get_sheet_by_name('lista')
# totallista = lista.rows
# for filas in totallista[:]:
#     a += 1
#     if a > 1:
#         periodo = Periodo.objects.get(nombre=filas[2].value)
#         filas[1].value = periodo.id
#         print('periodo actualizado en la fila ' + str(a))
# miarchivo.save("listaestudiantes_todos.xlsx")
# a = 0
# miarchivo = openpyxl.load_workbook('listaestudiantes_todos.xlsx')
# lista = miarchivo.get_sheet_by_name('lista')
# totallista = lista.rows
# for filas in totallista[:]:
#     a += 1
#     if a > 1:
#         if not Nivel.objects.filter(periodo__id=filas[1].value,sesion__id=filas[14].value,nivellibrecoordinacion__coordinacion__id=filas[6].value).exists():
#             nivel =Nivel(periodo_id=filas[1].value,
#                   sesion_id=filas[14].value,
#                   inicio=datetime.now().date(),
#                   fin=datetime.now().date(),
#                   paralelo='',
#                   modalidad_id=1,
#                   sede_id=1,
#                   cerrado=True,
#                   fechatopematricula=datetime.now().date(),
#                   fechatopematriculaex=datetime.now().date(),
#                   fechatopematriculaes=datetime.now().date(),
#                   nivelgrado=False,
#                   aplicabecas=True)
#             nivel.save()
#             coordinacion = Coordinacion.objects.get(pk=filas[6].value)
#             nivel.coordinacion(coordinacion)
#             print('ingresando nivel ' + str(a))
#
#
#
#
# # /*actualiza nivel*/
# # import openpyxl
# a = 0
# miarchivo = openpyxl.load_workbook('listaestudiantes_todos.xlsx')
# lista = miarchivo.get_sheet_by_name('lista')
# totallista = lista.rows
# for filas in totallista[:]:
#     a += 1
#     if a > 1:
#         if Nivel.objects.filter(periodo__id=filas[1].value, sesion__id=filas[14].value, nivellibrecoordinacion__coordinacion__id=filas[6].value).exists():
#             niv = Nivel.objects.get(periodo__id=filas[1].value, sesion__id=filas[14].value, nivellibrecoordinacion__coordinacion__id=filas[6].value)
#             filas[15].value = niv.id
#             print('actualizando nivel en excell ' + str(a))
# miarchivo.save("listaestudiantes_todos.xlsx")
#
# # /*ingresa matriculas*/
# # import openpyxl
# a = 0
# miarchivo = openpyxl.load_workbook('listaestudiantes_todos.xlsx')
# lista = miarchivo.get_sheet_by_name('lista')
# totallista = lista.rows
# for filas in totallista[:]:
#     a += 1
#     if a > 1:
#         # if Nivel.objects.filter(periodo__id=filas[1].value,sesion__id=filas[14].value,nivellibrecoordinacion__coordinacion__id=filas[6].value).exists():
#         if not Matricula.objects.filter(inscripcion_id=filas[11].value, nivel_id=filas[15].value).exists():
#             matri = Matricula(inscripcion_id=filas[11].value,
#                                           nivel_id=filas[15].value,
#                                           nivelmalla_id=filas[8].value,
#                                           estado_matricula__in=[2,3],
#                                           tipomatricula_id=1,
#                                           cerrada=True,
#                                           paralelo_id=filas[10].value)
#             matri.save()
#             filas[16].value = 'SI'
#             print('Matriculando estudiante ' + str(matri))
# miarchivo.save("listaestudiantes_todos.xlsx")

# import openpyxl
# a = 0
# miarchivo = openpyxl.load_workbook('listaestudiantes_todos.xlsx')
# lista = miarchivo.get_sheet_by_name('lista')
# totallista = lista.rows
# for filas in totallista[:]:
#     a += 1
#     if a > 1:
#         # if Nivel.objects.filter(periodo__id=filas[1].value,sesion__id=filas[14].value,nivellibrecoordinacion__coordinacion__id=filas[6].value).exists():
#             if Matricula.objects.filter(inscripcion_id=filas[11].value, nivel__periodo__id=filas[1].value).count()==1:
#                 matri = Matricula.objects.get(inscripcion_id=filas[11].value, nivel__periodo__id=filas[1].value)
#                 filas[17].value = matri.id
#                 matri.nivelmalla_id = filas[8].value
#                 matri.save()
#                 print('Matriculando estudiante ' + str(a))
#             else: filas[18].value = 'dos veces'
# miarchivo.save("listaestudiantes_todos.xlsx")

# lecciones=Leccion.objects.filter(status=True, fecha__gte='2017-12-04', fecha__lte='2017-12-10').exclude(clase__materia__asignaturamalla__malla__carrera__coordinacion__id__in=[5,7,9]).order_by('clase__materia__asignaturamalla__malla__carrera__coordinacion','fecha')
# n=1
# for leccion in lecciones:
#     print(n)
#     n+=1
#     print("fecha: %s, coordinacion: %s materia: %s" % (leccion.fecha,leccion.clase.materia.asignaturamalla.malla.carrera.coordinacion_set.filter(status=True)[0].nombre,leccion.clase.materia.asignatura))
#     leccion.delete()

# for materiaasignada in MateriaAsignada.objects.filter(status=True, matricula__nivel__periodo__id=14).exclude(materia__asignaturamalla__malla__carrera__coordinacion__id__in=[5,7,9]).order_by('materia__asignaturamalla__malla__carrera__coordinacion'):
#     print(materiaasignada)
#     materiaasignada.actualiza_notafinal()
# print("listo")


# arreglar asistencias
# fechas = ['2017-10-07','2017-10-08','2017-10-14','2017-10-15']
#
# for materia in Materia.objects.filter(status=True, nivel__periodo__id=14, asignaturamalla__malla__carrera__coordinacion=9).order_by('profesormateria__profesor','id'):
#     for fechalista in fechas:
#         dia = convertirfecha2(fechalista).isoweekday()
#         bandera = 0
#
#         for turno in Turno.objects.filter(clase__materia=materia, clase__dia=dia, clase__activo=True).distinct():
#             bandera = 0
#             cl = materia.clase_set.filter(inicio__lte=fechalista, fin__gte=fechalista, dia=dia, activo=True,turno=turno)
#             if cl:
#                 if not Leccion.objects.filter(fecha=fechalista, clase__materia=materia, clase__turno=turno, clase__materia__profesormateria__profesor=materia.profesor_principal()).exists():
#                     bandera = 1
#                     print ("materia: %s, profesor: %s, fecha: %s" % (materia,materia.profesor_principal(),fechalista))
#                     copia_clases(cl[0] , fechalista)

# contador=0
# g = Group.objects.get(pk=2)
# for inscripcion in Inscripcion.objects.all().exclude(persona__usuario__groups__id=2):
#     contador += 1
#     g.user_set.add(inscripcion.persona.usuario)
#     g.save()
#     print(contador.__str__()+" "+inscripcion.__str__())

# contador=0
# for leccion in Leccion.objects.filter(horasalida__isnull=True, fecha__lt='2018-11-07'):
#     leccion.horasalida = leccion.clase.turno.termina
#     leccion.abierta = False
#     leccion.save()
#     contador += 1
#     print(contador.__str__() + " " + leccion.__str__())

# MASIVO ENVIO DE CORREO DE CREDENCIALES A ADMISION
# workbook = xlrd.open_workbook("actualizarcorreoadmision.xlsx")
# sheet = workbook.sheet_by_index(0)
# linea = 1
# cuenta = 20
# from django.db import transaction
# for rowx in range(sheet.nrows):
#     if linea>1:
#         cols = sheet.row_values(rowx)
#         cedula = cols[0].__str__().strip()
#         correo = cols[1].__str__().lower().strip()
#         if not cedula or not correo:
#             print('No existe cedula o correo %s - %s' % (cedula, correo))
#         elif not Inscripcion.objects.filter(persona__cedula=cedula, modalidad_id=3).exists():
#             print ('Inscripcion en la modalidad no existe %s' % cedula)
#         else:
#             inscripcion = Inscripcion.objects.filter(persona__cedula=cedula, modalidad_id=3)
#             # if inscripcion.values('id').filter(persona__email=correo).exists():
#             try:
#                 inscripcion = inscripcion[0]
#                 persona = inscripcion.persona
#                 persona.email = correo
#                 persona.save()
#                 linea += 1
#                 inscripcion.envio_correo_admision(cuenta)
#                 print('%s - Datos procesados %s - %s - %s - %s' % (linea.__str__(), cedula, correo, inscripcion.__str__(), inscripcion.id.__str__()))
#             except Exception as ex:
#                 transaction.set_rollback(True)
#                 print('Error al procesar los datos %s - %s' % (cedula, correo))
#             # else:
#             #     print('Ya existe correo en la inscripcion %s - %s' % (cedula, correo))
#     else:
#         linea += 1


# workbook = xlrd.open_workbook("sacarcedula.xlsx")
# sheet = workbook.sheet_by_index(0)
# linea = 1
# cuenta = 20
# from django.db import transaction
# for rowx in range(sheet.nrows):
#     if linea>1:
#         cols = sheet.row_values(rowx)
#         apellidos = cols[0].__str__().strip().split(" ")
#         nombres = cols[1].__str__().strip()
#         correo = cols[3].__str__().strip()
#         if Persona.objects.filter(apellido1=apellidos[0], apellido2=apellidos[1], nombres=nombres):
#             persona = Persona.objects.filter(apellido1=apellidos[0], apellido2=apellidos[1], nombres=nombres)[0]
#             print('%s - %s - %s' % (persona.cedula, correo, persona.nombre_completo_inverso()))
#         else:
#             print('No existe %s - %s' % (cols[0], cols[1]))
#     else:
#         linea += 1
