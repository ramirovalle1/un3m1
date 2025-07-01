# -*- coding: UTF-8 -*-
import random
import os
import io
import sys
import fitz
import uuid
from datetime import datetime, timedelta, date
import time
import pyqrcode
from django.contrib import messages
from django.db.models import Sum
import xlwt
from django.db.models import Avg
from django.core.files import File as DjangoFile
from xlwt import *
from django.contrib.auth.decorators import login_required
from django.db import transaction, connection, connections
from django.db.models.query_utils import Q
from django.http import HttpResponseRedirect, HttpResponse, JsonResponse
from django.shortcuts import render

from core.firmar_documentos import obtener_posicion_x_y_saltolinea, firmarmasivo, obtener_posicion_x_y_saltolinea_firmapagina
from core.firmar_documentos_ec import JavaFirmaEc
from decorators import secure_module, last_access
from inno.models import TipoActaFirma
from sagest.models import PersonaDepartamentoFirmas, DistributivoPersona, DistributivoPersonaHistorial
from sga.commonviews import adduserdata, traerNotificaciones
from sga.forms import GraduadoForm, ActaFacultadForm, ActaGraduadosForm, GraduadoIntegracionCurricularForm
from sga.funciones import MiPaginador, log, puede_realizar_accion, resetear_clave, remover_caracteres_tildes_unicode, \
    null_to_decimal, generar_nombre, variable_valor, puede_realizar_accion_afirmativo, remover_caracteres_especiales_unicode
from sga.funciones_templatepdf import actatribunalcalificacion, rubricatribunalcalificacion, actagradoposgrado, \
    actagradoposgradograduados, actatitulacioncomplexivo, actagradocomplexivo, actagradocomplexivofirma, actaconsolidadafirma, getactagradoposgradograduados2
from sga.models import Inscripcion, Graduado, Carrera, Egresado, PerfilInscripcion, \
    RecordAcademico, Profesor, PracticasPreprofesionalesInscripcion, Persona, CargoInstitucion, \
    null_to_numeric, ParticipantesMatrices, ItemExamenComplexivo, ExamenComlexivoGraduados, AsignaturaMalla, \
    Administrativo, Sexo, Coordinacion, ActaFacultad, MatriculaTitulacion, ComplexivoGrupoTematica, \
    ComplexivoDetalleGrupo, RubricaTitulacion, TemaTitulacionPosgradoMatricula, TribunalTemaTitulacionPosgradoMatricula, \
    TipoActa, TIPO_ACTAGRADUADO, FirmaPersona, CUENTAS_CORREOS, Notificacion, MateriaTitulacion, CertificadoIdioma, \
    ResponsableCoordinacion, Malla, Periodo
from sga.funcionesxhtml2pdf import conviert_html_to_pdf, conviert_html_to_pdfsaveqrcertificado, \
    conviert_html_to_pdfsaveqrtitulo
from pdf2image import convert_from_bytes
from django.db.models.functions import ExtractYear
from django.template.loader import get_template
from settings import SITE_STORAGE, PUESTO_ACTIVO_ID, SITE_POPPLER, DEBUG
from sga.tasks import send_html_mail
import openpyxl
from sga.excelbackground import reporte_tituloinsigniamasivo_background, export_reportegraduados2

mes = {1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo", 6: "Junio",
       7: "Julio", 8: "Agosto", 9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"}

MONEDA_SINGULAR = 'dolar'
MONEDA_PLURAL = 'dolares'
CENTIMOS_SINGULAR = 'centavo'
CENTIMOS_PLURAL = 'centavos'
MAX_NUMERO = 999999999999
UNIDADES = (
    'cero',
    'uno',
    'dos',
    'tres',
    'cuatro',
    'cinco',
    'seis',
    'siete',
    'ocho',
    'nueve'
)

DECENAS = (
    'diez',
    'once',
    'doce',
    'trece',
    'catorce',
    'quince',
    u'dieciséis',
    'diecisiete',
    'dieciocho',
    'diecinueve'
)

DIEZ_DIEZ = (
    'cero',
    'diez',
    'veinte',
    'treinta',
    'cuarenta',
    'cincuenta',
    'sesenta',
    'setenta',
    'ochenta',
    'noventa'
)

CIENTOS = (
    '_',
    'ciento',
    'doscientos',
    'trescientos',
    'cuatroscientos',
    'quinientos',
    'seiscientos',
    'setecientos',
    'ochocientos',
    'novecientos'
)


def fecha_letra(valor):
    mes = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    a = int(valor[0:4])
    m = int(valor[5:7])
    d = int(valor[8:10])
    if d == 1:
        return u"al %s día del mes de %s del %s" % (numero_a_letras(d), str(mes[m - 1]), numero_a_letras(a))
    else:
        if d == 21:
            return u"a los %s días del mes de %s del %s" % (u'veintiún', str(mes[m - 1]), numero_a_letras(a))
        else:
            if d == 31:
                return u"a los %s días del mes de %s del %s" % (u'treintaiún', str(mes[m - 1]), numero_a_letras(a))
            else:
                return u"a los %s días del mes de %s del %s" % (numero_a_letras(d), str(mes[m - 1]), numero_a_letras(a))


def numero_a_letras(numero):
    numero_entero = int(numero)
    if numero_entero > MAX_NUMERO:
        raise OverflowError('Número demasiado alto')
    if numero_entero < 0:
        return 'menos %s' % numero_a_letras(abs(numero))
    letras_decimal = ''
    parte_decimal = int(null_to_decimal(((abs(numero) - abs(numero_entero)) * 100), 1))
    if parte_decimal > 9:
        letras_decimal = 'punto %s' % numero_a_letras(parte_decimal)
    elif parte_decimal > 0:
        letras_decimal = 'punto cero %s' % numero_a_letras(parte_decimal)
    if numero_entero <= 99:
        resultado = leer_decenas(numero_entero)
    elif numero_entero <= 999:
        resultado = leer_centenas(numero_entero)
    elif numero_entero <= 999999:
        resultado = leer_miles(numero_entero)
    elif numero_entero <= 999999999:
        resultado = leer_millones(numero_entero)
    else:
        resultado = leer_millardos(numero_entero)
    resultado = resultado.replace('uno mil', 'un mil')
    resultado = resultado.strip()
    resultado = resultado.replace(' _ ', ' ')
    resultado = resultado.replace('  ', ' ')
    if parte_decimal > 0:
        resultado = '%s %s' % (resultado, letras_decimal)
    return resultado


def leer_decenas(numero):
    if numero < 10:
        return UNIDADES[numero]
    decena, unidad = divmod(numero, 10)
    if unidad == 0:
        resultado = DIEZ_DIEZ[decena]
    else:
        if numero <= 19:
            resultado = DECENAS[unidad]
        elif numero <= 29:
            resultado = 'veinti%s' % UNIDADES[unidad]
        else:
            resultado = DIEZ_DIEZ[decena]
            if unidad > 0:
                resultado = '%s y %s' % (resultado, UNIDADES[unidad])
    return resultado


def leer_centenas(numero):
    centena, decena = divmod(numero, 100)
    if centena == 1:
        resultado = 'cien'
    else:
        if numero == 0:
            resultado = 'cien'
        else:
            resultado = CIENTOS[centena]
            if decena > 0:
                resultado = '%s %s' % (resultado, leer_decenas(decena))
    return resultado


def leer_miles(numero):
    millar, centena = divmod(numero, 1000)
    resultado = ''
    if millar == 1:
        resultado = ''
    if 2 <= millar <= 9:
        resultado = UNIDADES[millar]
    elif 10 <= millar <= 99:
        resultado = leer_decenas(millar)
    elif 100 <= millar <= 999:
        resultado = leer_centenas(millar)
    resultado = '%s mil' % resultado
    if centena > 0:
        resultado = '%s %s' % (resultado, leer_centenas(centena))
    return resultado


def leer_millones(numero):
    millon, millar = divmod(numero, 1000000)
    resultado = ''
    if millon == 1:
        resultado = ' un millon '
    if 2 <= millon <= 9:
        resultado = UNIDADES[millon]
    elif 10 <= millon <= 99:
        resultado = leer_decenas(millon)
    elif 100 <= millon <= 999:
        resultado = leer_centenas(millon)
    if millon > 1:
        resultado = '%s millones' % resultado
    if (millar > 0) and (millar <= 999):
        resultado = '%s %s' % (resultado, leer_centenas(millar))
    elif (millar >= 1000) and (millar <= 999999):
        resultado = '%s %s' % (resultado, leer_miles(millar))
    return resultado


def leer_millardos(numero):
    millardo, millon = divmod(numero, 1000000)
    return '%s millones %s' % (leer_miles(millardo), leer_millones(millon))


@login_required(redirect_field_name='ret', login_url='/loginsga')
@secure_module
@last_access
@transaction.atomic()
def view(request):
    data = {}
    adduserdata(request, data)
    data['persona'] = persona = request.session['persona']
    periodo = request.session['periodo']
    miscarreras = persona.mis_carreras()
    horaactual = datetime.now().strftime('%Y%m%d_%H%M%S')
    user = request.user
    dominio_sistema = 'https://sga.unemi.edu.ec'
    if DEBUG:
        dominio_sistema = 'http://localhost:8000'
    data["DOMINIO_DEL_SISTEMA"] = dominio_sistema

    data['IS_DEBUG'] = IS_DEBUG = variable_valor('IS_DEBUG')

    if request.method == 'POST':
        action = request.POST['action']

        if action == 'add':
            try:
                puede_realizar_accion(request, 'sga.puede_adicionar_graduado')
                inscripcion = Inscripcion.objects.get(pk=request.POST['id'])
                f = GraduadoForm(request.POST)
                if f.is_valid():
                    directorcarrera = None
                    representantedocente = None
                    representanteestudiantil = None

                    representanteservidores = None
                    representantesuplentedocente = None
                    representantesuplenteestudiantil = None
                    representantesuplenteservidores = None
                    docentesecretario = None
                    integrantetribunal = None
                    datosprofesor = None
                    directoresfacultad = None

                    if not Graduado.objects.filter(inscripcion=inscripcion).exists():
                        subdecano = None
                        if request.POST['decano']:
                            if int(request.POST['decano']) > 0:
                                decano = Profesor.objects.get(pk=int(request.POST['decano']))

                        if request.POST['subdecano']:
                            if int(request.POST['subdecano']) > 0:
                                subdecano = Profesor.objects.get(pk=int(request.POST['subdecano']))

                        if request.POST['representantedocente'] and request.POST['representantedocente'] != '0':
                            representantedocente = Profesor.objects.get(pk=int(request.POST['representantedocente']))

                        if request.POST['representantesuplentedocente'] and request.POST['representantesuplentedocente'] != '0':
                            representantesuplentedocente = Profesor.objects.get(pk=int(request.POST['representantesuplentedocente']))

                        if 'docenteseretario' in request.POST:
                            if request.POST['docenteseretario'] and request.POST['docenteseretario'] != '0':
                                docenteseretario = Profesor.objects.get(pk=int(request.POST['docenteseretario']))

                        if 'integrantetribunal' in request.POST:
                            if request.POST['integrantetribunal'] and request.POST['integrantetribunal'] != '0':
                                integrantetribunal = Profesor.objects.get(pk=int(request.POST['integrantetribunal']))

                        if 'profesor' in request.POST:
                            if request.POST['profesor'] and request.POST['profesor'] != '0':
                                datosprofesor = Profesor.objects.get(pk=int(request.POST['profesor']))



                        if request.POST['representanteestudiantil'] and request.POST['representanteestudiantil'] != '0':
                            representanteestudiantil = int(request.POST['representanteestudiantil'])

                        if request.POST['representantesuplenteestudiantil'] and request.POST['representantesuplenteestudiantil'] != '0':
                            representantesuplenteestudiantil = int(request.POST['representantesuplenteestudiantil'])

                        if request.POST['representanteservidores'] and request.POST['representanteservidores'] != '0':
                            representanteservidores = Administrativo.objects.get(persona_id=int(request.POST['representanteservidores']))

                        if request.POST['representantesuplenteservidores'] and request.POST['representantesuplenteservidores'] != '0':
                            representantesuplenteservidores = Administrativo.objects.get(pk=int(request.POST['representantesuplenteservidores']))

                        if 'directorcarrera' in request.POST:
                            if int(request.POST['directorcarrera']) > 0:
                                directorcarrera = Profesor.objects.get(pk=int(request.POST['directorcarrera']))
                        graduado = Graduado(inscripcion=inscripcion,
                                            tematesis=f.cleaned_data['tematesis'],
                                            notafinal=inscripcion.datos_egresado().notaegreso,
                                            fechagraduado=f.cleaned_data['fechagraduado'],
                                            registro=f.cleaned_data['registro'],
                                            numeroactagrado=f.cleaned_data['numeroactagrado'],
                                            promediotitulacion=f.cleaned_data['promediotitulacion'],
                                            fechaactagrado=f.cleaned_data['fechaactagrado'],
                                            fecharefrendacion=f.cleaned_data['fecharefrendacion'] if f.cleaned_data['fecharefrendacion'] else None,
                                            # mecanismotitulacion=f.cleaned_data['mecanismotitulacion'],
                                            codigomecanismotitulacion=f.cleaned_data['codigomecanismotitulacion'],
                                            fechaconsejo=f.cleaned_data['fechaconsejo'],
                                            nombretitulo=f.cleaned_data['nombretitulo'],
                                            # profesor_id=f.cleaned_data['profesor'] if int(f.cleaned_data['profesor'])>0 else None,
                                            # profesor_id=f.cleaned_data['profesor'] if request.POST['profesor'] else None,
                                            # integrantetribunal_id=f.cleaned_data['integrantetribunal'] if request.POST['integrantetribunal'] else None,
                                            # docentesecretario_id= f.cleaned_data['docentesecretario'] if request.POST['docentesecretario'] else None,
                                            profesor_id=request.POST['profesor'] if datosprofesor else None,
                                            integrantetribunal_id=request.POST['integrantetribunal'] if integrantetribunal else None,
                                            docentesecretario_id=request.POST['docentesecretario'] if docentesecretario else None,
                                            asistentefacultad_id=f.cleaned_data['asistentefacultad'] if int(f.cleaned_data['asistentefacultad']) > 0 else None,
                                            secretariageneral_id=f.cleaned_data['secretariageneral'] if int(f.cleaned_data['secretariageneral']) > 0 else None,
                                            horagraduacion=f.cleaned_data['horagraduacion'],
                                            horacertificacion=f.cleaned_data['horacertificacion'],
                                            folio=f.cleaned_data['folio'],
                                            horastitulacion=f.cleaned_data['horastitulacion'],
                                            creditotitulacion=f.cleaned_data['creditotitulacion'],
                                            creditovinculacion=f.cleaned_data['creditovinculacion'],
                                            creditopracticas=f.cleaned_data['creditopracticas'],
                                            notagraduacion=f.cleaned_data['notagraduacion'],
                                            decano_id=decano.persona.id if request.POST['decano'] else None,
                                            subdecano_id=subdecano.persona.id if subdecano else None,  # request.POST['subdecano']
                                            representanteestudiantil_id=request.POST['representanteestudiantil'] if representanteestudiantil else None,
                                            representantesuplenteestudiantil_id=request.POST['representantesuplenteestudiantil'] if representantesuplenteestudiantil else None,
                                            representantedocente_id=representantedocente.persona.id if representantedocente else None,  # request.POST['representantedocente']
                                            representantesuplentedocente_id=representantesuplentedocente.persona.id if representantesuplentedocente else None,  # request.POST['representantedocente']
                                            representanteservidores_id=representanteservidores.persona.id if representanteservidores else None,
                                            representantesuplenteservidores_id=representantesuplenteservidores.persona.id if representantesuplenteservidores else None,
                                            estadograduado=f.cleaned_data['estadograduado'])
                        graduado.save(request)
                        if 'directoresfacultad' in request.POST:
                            for dato in f.cleaned_data['directoresfacultad']:
                                graduado.directoresfacultad.add(dato)
                                graduado.save()
                        # graduado.promediotitulacion = null_to_decimal((graduado.promediogrado + graduado.notatesis + graduado.sustentacion) / 3, 2)
                        # if graduado.vinculacion == 0 and graduado.practicas == 0:
                        #     graduado.notagraduacion = null_to_decimal((graduado.notafinal + graduado.promediotitulacion + graduado.pasantias) / 3, 2)
                        # elif graduado.vinculacion != 0 and graduado.practicas != 0:
                        #     graduado.notagraduacion = null_to_decimal((graduado.notafinal + graduado.promediotitulacion + graduado.pasantias + graduado.vinculacion + graduado.practicas) / 5, 2)
                        # elif graduado.vinculacion != 0:
                        #     graduado.notagraduacion = null_to_decimal((graduado.notafinal + graduado.promediotitulacion + graduado.pasantias + graduado.vinculacion) / 4, 2)
                        # elif graduado.practicas != 0:
                        #     graduado.notagraduacion = null_to_decimal((graduado.notafinal + graduado.promediotitulacion + graduado.pasantias + graduado.practicas) / 4, 2)
                        # graduado.save()
                        log(u'Adiciono graduado: %s nota: %s' % (graduado, str(graduado.notafinal)), request, "add")
                        if f.cleaned_data['fechaegresadoactivo']:
                            egresado = Egresado.objects.get(inscripcion_id=graduado.inscripcion.id)
                            egresado.fechaegreso = f.cleaned_data['fechaegresado']
                            egresado.save(request)
                            log(u'Modifico fecha de egreso: %s' % egresado, request, "edit")
                        if f.cleaned_data['fechainicioactivo']:
                            inscripcion = Inscripcion.objects.get(pk=graduado.inscripcion.id)
                            inscripcion.fechainicioprimernivel = f.cleaned_data['fechainicio']
                            inscripcion.save(request)
                            log(u'Modifico fecha de inicio de primer Niver: %s' % inscripcion, request, "edit")
                        return JsonResponse({"result": "ok", "id": graduado.id})
                    else:
                        return JsonResponse({"result": "bad", "mensaje": u"Ya el estudiante esta graduado."})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos. %" % ex.__str__()})

        elif action == 'edit':
            try:
                graduado = Graduado.objects.get(pk=request.POST['id'])
                f = GraduadoForm(request.POST)
                if f.is_valid():
                    graduado.directoresfacultad.clear()
                    falta = ''
                    directorcarrera = None
                    subdecano = None

                    representantedocente = None
                    representantedocentedos = None
                    representantesuplentedocente = None
                    representantesuplentedocentedos = None
                    representanteestudiantil = None
                    representanteestudiantildos = None
                    representantesuplenteestudiantil = None
                    representantesuplenteestudiantildos = None
                    representanteservidores = None
                    representanteservidoresdos = None
                    representantesuplenteservidores = None
                    representantesuplenteservidoresdos = None
                    directoresfacultad = None

                    noaplicapropuesta = False
                    if graduado.matriculatitulacion:
                        if graduado.matriculatitulacion.alternativa.aplicapropuesta:
                            noaplicapropuesta = True
                    if noaplicapropuesta:
                        if not int(request.POST['docenteevaluador1']) > 0:
                            falta = "Evaluador 1, "
                        if not int(request.POST['docenteevaluador2']) > 0:
                            falta = "Evaluador 2, "
                    else:
                        if not request.POST['integrantetribunal'] or request.POST['integrantetribunal'] == '0':
                            falta = "integrante del tribunal, "
                        if not request.POST['docentesecretario'] or request.POST['docentesecretario'] == '0':
                            falta += "docente secretario, "
                        if not request.POST['profesor'] or request.POST['profesor'] == '0':
                            falta += "tutor del trabajo,"

                    # if not request.POST['representanteestudiantil'] or request.POST['representanteestudiantil']=='0':
                    #     if not request.POST['representantesuplenteestudiantil'] or request.POST['representantesuplenteestudiantil']=='0':
                    #         falta += "representante principal o suplente estudiantil, "
                    #
                    #
                    # if not request.POST['representantedocente'] or request.POST['representantedocente']=='0':
                    #     if not request.POST['representantesuplentedocente'] or request.POST['representantesuplentedocente'] == '0':
                    #         falta += "representante principal o suplente de docente, "
                    #
                    #
                    # if not request.POST['representanteservidores'] or request.POST['representanteservidores']=='0':
                    #     if not request.POST['representantesuplenteservidores'] or request.POST['representantesuplenteservidores'] == '0':
                    #         falta += "representante principal o suplente de servidores y trabajadores, "

                    if falta != '':
                        return JsonResponse({"result": "bad", "mensaje": u"Complete los campos: " + falta[:falta.__len__() - 2]})

                    if int(request.POST['decano']) > 0:
                        decano = Profesor.objects.get(pk=int(request.POST['decano']))
                    if request.POST['subdecano']:
                        if int(request.POST['subdecano']) > 0:
                            subdecano = Profesor.objects.get(pk=int(request.POST['subdecano']))

                    if request.POST['representantedocente'] and request.POST['representantedocente'] != '0':
                        representantedocente = Profesor.objects.get(pk=int(request.POST['representantedocente']))

                    if request.POST['representantedocentedos'] and request.POST['representantedocentedos'] != '0':
                        representantedocentedos = Profesor.objects.get(pk=int(request.POST['representantedocentedos']))

                    if request.POST['representantesuplentedocente'] and request.POST['representantesuplentedocente'] != '0':
                        representantesuplentedocente = Profesor.objects.get(pk=int(request.POST['representantesuplentedocente']))

                    if request.POST['representantesuplentedocentedos'] and request.POST['representantesuplentedocentedos'] != '0':
                        representantesuplentedocentedos = Profesor.objects.get(pk=int(request.POST['representantesuplentedocentedos']))

                    if request.POST['representanteestudiantil'] and request.POST['representanteestudiantil'] != '0':
                        representanteestudiantil = int(request.POST['representanteestudiantil'])

                    if request.POST['representanteestudiantildos'] and request.POST['representanteestudiantildos'] != '0':
                        representanteestudiantildos = int(request.POST['representanteestudiantildos'])

                    if request.POST['representantesuplenteestudiantil'] and request.POST['representantesuplenteestudiantil'] != '0':
                        representantesuplenteestudiantil = int(request.POST['representantesuplenteestudiantil'])

                    if request.POST['representantesuplenteestudiantildos'] and request.POST['representantesuplenteestudiantildos'] != '0':
                        representantesuplenteestudiantildos = int(request.POST['representantesuplenteestudiantildos'])

                    if request.POST['representanteservidores'] and request.POST['representanteservidores'] != '0':
                        representanteservidores = Administrativo.objects.get(pk=int(request.POST['representanteservidores']))

                    if request.POST['representanteservidoresdos'] and request.POST['representanteservidoresdos'] != '0':
                        representanteservidoresdos = Administrativo.objects.get(pk=int(request.POST['representanteservidoresdos']))

                    if request.POST['representantesuplenteservidores'] and request.POST['representantesuplenteservidores'] != '0':
                        representantesuplenteservidores = Administrativo.objects.get(pk=int(request.POST['representantesuplenteservidores']))

                    if request.POST['representantesuplenteservidoresdos'] and request.POST['representantesuplenteservidoresdos'] != '0':
                        representantesuplenteservidoresdos = Administrativo.objects.get(pk=int(request.POST['representantesuplenteservidoresdos']))

                    if 'directoresfacultad' in request.POST:
                        directoresfacultad = f.cleaned_data['directoresfacultad']
                        graduado.directoresfacultad.clear()
                        for df in directoresfacultad:
                            graduado.directoresfacultad.add(df)

                    if 'directorcarrera' in request.POST:
                        if int(request.POST['directorcarrera']) > 0:
                            directorcarrera = Profesor.objects.get(pk=int(request.POST['directorcarrera']))

                    graduado.tematesis = f.cleaned_data['tematesis']
                    graduado.fechagraduado = f.cleaned_data['fechagraduado']
                    graduado.registro = f.cleaned_data['registro']
                    graduado.numeroactagrado = f.cleaned_data['numeroactagrado']
                    graduado.fechaactagrado = f.cleaned_data['fechaactagrado']
                    graduado.fecharefrendacion = f.cleaned_data['fecharefrendacion'] if f.cleaned_data['fecharefrendacion'] else None
                    graduado.mecanismotitulacion = ''
                    graduado.codigomecanismotitulacion = f.cleaned_data['codigomecanismotitulacion']
                    graduado.fechaconsejo = f.cleaned_data['fechaconsejo']
                    graduado.nombretitulo = f.cleaned_data['nombretitulo']
                    graduado.asistentefacultad_id = f.cleaned_data['asistentefacultad'] if int(f.cleaned_data['asistentefacultad']) > 0 else None
                    graduado.secretariageneral_id = f.cleaned_data['secretariageneral'] if int(f.cleaned_data['secretariageneral']) > 0 else None
                    graduado.horagraduacion = f.cleaned_data['horagraduacion']
                    graduado.horacertificacion = f.cleaned_data['horacertificacion']
                    graduado.folio = f.cleaned_data['folio']
                    graduado.promediotitulacion = f.cleaned_data['promediotitulacion']
                    graduado.horastitulacion = f.cleaned_data['horastitulacion']
                    graduado.creditotitulacion = f.cleaned_data['creditotitulacion']
                    graduado.creditovinculacion = f.cleaned_data['creditovinculacion']
                    graduado.creditopracticas = f.cleaned_data['creditopracticas']
                    graduado.periodo = periodo
                    graduado.decano_id = decano.persona.id if int(request.POST['decano']) > 0 else None
                    graduado.subdecano_id = subdecano.persona.id if subdecano else None

                    graduado.representanteestudiantil_id = request.POST['representanteestudiantil'] if representanteestudiantil else None
                    graduado.representanteestudiantildos_id = request.POST['representanteestudiantildos'] if representanteestudiantildos else None

                    graduado.representantesuplenteestudiantil_id = request.POST['representantesuplenteestudiantil'] if representantesuplenteestudiantil else None
                    graduado.representantesuplenteestudiantildos_id = request.POST['representantesuplenteestudiantildos'] if representantesuplenteestudiantildos else None

                    graduado.representantedocente_id = representantedocente.persona.id if representantedocente else None
                    graduado.representantedocentedos_id = representantedocentedos.persona.id if representantedocentedos else None

                    graduado.representantesuplentedocente_id = representantesuplentedocente.persona.id if representantesuplentedocente else None
                    graduado.representantesuplentedocentedos_id = representantesuplentedocentedos.persona.id if representantesuplentedocentedos else None

                    graduado.representanteservidores_id = representanteservidores.persona.id if representanteservidores else None
                    graduado.representanteservidoresdos_id = representanteservidoresdos.persona.id if representanteservidoresdos else None

                    graduado.representantesuplenteservidores_id = representantesuplenteservidores.persona.id if representantesuplenteservidores else None
                    graduado.representantesuplenteservidoresdos_id = representantesuplenteservidoresdos.persona.id if representantesuplenteservidoresdos else None

                    graduado.notafinal = graduado.inscripcion.promedio_record()
                    graduado.estadograduado = f.cleaned_data['estadograduado']
                    graduado.notagraduacion = f.cleaned_data['notagraduacion']
                    if noaplicapropuesta:
                        graduado.docenteevaluador1_id = f.cleaned_data['docenteevaluador1'] if int(f.cleaned_data['docenteevaluador1']) > 0 else None
                        graduado.docenteevaluador2_id = f.cleaned_data['docenteevaluador2'] if int(f.cleaned_data['docenteevaluador2']) > 0 else None
                        graduado.directorcarrera_id = directorcarrera.persona.id if int(request.POST['directorcarrera']) > 0 else None
                    else:
                        graduado.profesor_id = f.cleaned_data['profesor'] if int(f.cleaned_data['profesor']) > 0 else None
                        graduado.integrantetribunal_id = f.cleaned_data['integrantetribunal'] if int(f.cleaned_data['integrantetribunal']) > 0 else None
                        graduado.docentesecretario_id = f.cleaned_data['docentesecretario'] if int(f.cleaned_data['docentesecretario']) > 0 else None
                    graduado.save(request)
                    log(u'Modifico graduado: %s' % graduado, request, "edit")
                    if f.cleaned_data['fechaegresadoactivo']:
                        egresado = Egresado.objects.get(inscripcion_id=graduado.inscripcion.id)
                        egresado.fechaegreso = f.cleaned_data['fechaegresado']
                        egresado.save(request)
                        log(u'Modifico fecha de egreso: %s' % egresado, request, "edit")
                    if f.cleaned_data['fechainicioactivo']:
                        inscripcion = Inscripcion.objects.get(pk=graduado.inscripcion.id)
                        inscripcion.fechainicioprimernivel = f.cleaned_data['fechainicio']
                        inscripcion.save(request)
                        log(u'Modifico fecha de inicio de primer Niver: %s' % inscripcion, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'editintegracioncurricular':
            try:
                graduado = Graduado.objects.get(pk=request.POST['id'])
                f = GraduadoIntegracionCurricularForm(request.POST)
                if f.is_valid():
                    directoresfacultad = None
                    if int(request.POST['decano']) > 0:
                        decano = Profesor.objects.get(pk=int(request.POST['decano']))
                    if 'directoresfacultad' in request.POST:
                        directoresfacultad = f.cleaned_data['directoresfacultad']
                        graduado.directoresfacultad.clear()
                        for df in directoresfacultad:
                            graduado.directoresfacultad.add(df)
                    graduado.tematesis = f.cleaned_data['tematesis']
                    graduado.fechagraduado = f.cleaned_data['fechagraduado']
                    graduado.registro = f.cleaned_data['registro']
                    graduado.numeroactagrado = f.cleaned_data['numeroactagrado']
                    graduado.fechaactagrado = f.cleaned_data['fechaactagrado']
                    graduado.fecharefrendacion = f.cleaned_data['fecharefrendacion'] if f.cleaned_data['fecharefrendacion'] else None
                    graduado.mecanismotitulacion = ''
                    graduado.codigomecanismotitulacion = f.cleaned_data['codigomecanismotitulacion']
                    graduado.fechaconsejo = f.cleaned_data['fechaconsejo']
                    graduado.nombretitulo = f.cleaned_data['nombretitulo']
                    graduado.asistentefacultad_id = f.cleaned_data['asistentefacultad'] if int(f.cleaned_data['asistentefacultad']) > 0 else None
                    graduado.secretariageneral_id = f.cleaned_data['secretariageneral'] if int(f.cleaned_data['secretariageneral']) > 0 else None
                    graduado.horagraduacion = f.cleaned_data['horagraduacion']
                    graduado.horacertificacion = f.cleaned_data['horacertificacion']
                    graduado.folio = f.cleaned_data['folio']
                    graduado.folioactaconsolidada = f.cleaned_data['folioactaconsolidada']
                    graduado.numeroactaconsolidada = f.cleaned_data['numeroactaconsolidada']
                    graduado.promediotitulacion = f.cleaned_data['promediotitulacion']
                    graduado.horastitulacion = f.cleaned_data['horastitulacion']
                    graduado.creditotitulacion = f.cleaned_data['creditotitulacion']
                    graduado.creditovinculacion = f.cleaned_data['creditovinculacion']
                    graduado.creditopracticas = f.cleaned_data['creditopracticas']
                    graduado.periodo = periodo
                    graduado.decano_id = decano.persona.id if int(request.POST['decano']) > 0 else None
                    graduado.notafinal = graduado.inscripcion.promedio_record()
                    graduado.estadograduado = f.cleaned_data['estadograduado']
                    graduado.notagraduacion = f.cleaned_data['notagraduacion']
                    graduado.save(request)
                    log(u'Modifico graduado: %s' % graduado, request, "edit")
                    if f.cleaned_data['fechaegresadoactivo']:
                        if not Egresado.objects.filter(inscripcion_id=graduado.inscripcion.id):
                            alumnoegresado = Egresado(inscripcion_id=graduado.inscripcion.id,
                                                      notaegreso=graduado.inscripcion.promedio_record(),
                                                      fechaegreso=f.cleaned_data['fechaegresado'])
                            alumnoegresado.save(request)
                        egresado = Egresado.objects.get(inscripcion_id=graduado.inscripcion.id)
                        egresado.fechaegreso = f.cleaned_data['fechaegresado']
                        egresado.save(request)
                        log(u'Modifico fecha de egreso: %s' % egresado, request, "edit")
                    if f.cleaned_data['fechainicioactivo']:
                        inscripcion = Inscripcion.objects.get(pk=graduado.inscripcion.id)
                        inscripcion.fechainicioprimernivel = f.cleaned_data['fechainicio']
                        inscripcion.save(request)
                        log(u'Modifico fecha de inicio de primer Niver: %s' % inscripcion, request, "edit")

                    if graduado.fechagraduado:
                        # solo los graduados mayor a la fecha de este periodo se ejecutan las actas de grado y consolidadas para firmar mediante el sistema
                        periodofirmae = Periodo.objects.get(pk=177)
                        if graduado.fechagraduado > periodofirmae.fin:
                            if not graduado.tipoacta_set.filter(status=True):
                                tipo = TipoActa(nombre='Acta de Grado', tipo=5, graduado=graduado)
                                tipo.save(request)
                                tipoactafirma = TipoActaFirma(tipoacta=tipo,
                                                              persona_id=graduado.asistentefacultad_id,
                                                              turnofirmar=True,
                                                              orden=1)
                                tipoactafirma.save(request)
                                tipoactafirma = TipoActaFirma(tipoacta=tipo,
                                                              persona_id=graduado.decano_id,
                                                              orden=2)
                                tipoactafirma.save(request)
                                tipoactafirma = TipoActaFirma(tipoacta=tipo,
                                                              persona_id=graduado.secretariageneral_id,
                                                              orden=3)
                                tipoactafirma.save(request)

                                tipo = TipoActa(nombre='Acta Consolidada', tipo=6, graduado=graduado)
                                tipo.save(request)
                                tipoactafirma = TipoActaFirma(tipoacta=tipo,
                                                              persona_id=graduado.asistentefacultad_id,
                                                              turnofirmar=True,
                                                              orden=1)
                                tipoactafirma.save(request)

                                for dirfac in graduado.directoresfacultad.all():
                                    dirfacultad = dirfac
                                tipoactafirma = TipoActaFirma(tipoacta=tipo,
                                                              persona=dirfacultad,
                                                              orden=2)
                                tipoactafirma.save(request)

                                tipoactafirma = TipoActaFirma(tipoacta=tipo,
                                                              persona_id=graduado.decano_id,
                                                              orden=3)
                                tipoactafirma.save(request)

                                tipoactafirma = TipoActaFirma(tipoacta=tipo,
                                                              persona_id=graduado.secretariageneral_id,
                                                              orden=4)
                                tipoactafirma.save(request)

                                archivotitulacionfirmada = graduado.materiatitulacion.archivotitulacionfirmada
                                tipo = TipoActa(nombre='Acta de Titulación', tipo=10, archivo=archivotitulacionfirmada, graduado=graduado, actafirmada=True)
                                tipo.save(request)

                                generaactagradofe = actagradocomplexivofirma(graduado.id)
                                generaactaconsolidadafe = actaconsolidadafirma(graduado.id)
                                if generaactagradofe:
                                    tipoac = TipoActa.objects.filter(tipo=5, graduado=graduado)[0]
                                    tipoac.archivo='qrcode/actatitulacion/fe_actagrado_' + str(graduado.id) + '.pdf'
                                    tipoac.save(request)

                                if generaactaconsolidadafe:
                                    tipoac = TipoActa.objects.filter(tipo=6, graduado=graduado)[0]
                                    tipoac.archivo='qrcode/actatitulacion/fe_actaconsolidada_' + str(graduado.id) + '.pdf'
                                    tipoac.save(request)

                            else:
                                if TipoActa.objects.filter(tipo=5, graduado=graduado, actafirmada=False):
                                    tipoactagrado = TipoActa.objects.filter(tipo=5, graduado=graduado)[0]
                                    if not TipoActaFirma.objects.filter(tipoacta=tipoactagrado, firmado=True):
                                        generaactagradofe = actagradocomplexivofirma(tipoactagrado.graduado.id)
                                if TipoActa.objects.filter(tipo=6, graduado=graduado, actafirmada=False):
                                    tipoactaconsolidada = TipoActa.objects.filter(tipo=6, graduado=graduado)[0]
                                    if not TipoActaFirma.objects.filter(tipoacta=tipoactaconsolidada, firmado=True):
                                        generaactaconsolidadafe = actaconsolidadafirma(tipoactaconsolidada.graduado.id)
                                if graduado.materiatitulacion.actatitulacionfirmada:
                                    archivotitulacionfirmada = graduado.materiatitulacion.archivotitulacionfirmada
                                    if TipoActa.objects.filter(tipo=10, graduado=graduado):
                                        tipoactatitulacion = TipoActa.objects.filter(tipo=10, graduado=graduado)[0]
                                        tipoactatitulacion.archivo = archivotitulacionfirmada
                                        tipoactatitulacion.actafirmada = True
                                        tipoactatitulacion.save(request)
                    return JsonResponse({"result": "ok"})
                else:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'calificarexamen':
            try:
                if ExamenComlexivoGraduados.objects.filter(graduado_id=request.POST['idgraduado'], itemexamencomplexivo_id=request.POST['idex']).exists():
                    detalle = ExamenComlexivoGraduados.objects.get(graduado_id=request.POST['idgraduado'], itemexamencomplexivo_id=request.POST['idex'])
                    detalle.examen = request.POST['valor']
                    notapon = float(request.POST['valor'])
                    detalle.ponderacion = null_to_decimal((notapon / 2), 2)
                    detalle.save(request)
                    # cursor = connections['sga_select'].cursor()
                    # sql="select null_to_decimal(("+str(notapon)+"/2),2)"
                    # cursor.execute(sql)
                    # results = cursor.fetchall()
                    # for r in results:
                    #     campo1 = r[0]
                    # detalle.ponderacion=campo1
                    # detalle.save(request)
                    # graduado = Graduado.objects.get(pk=request.POST['id'])
                    log(u'Adicionó Examen Complexivo graduado: %s' % detalle, request, "add")
                else:
                    # cursor = connections['sga_select'].cursor()
                    # sql = "select null_to_decimal((" + str(request.POST['valor']) + "/2),2)"
                    # cursor.execute(sql)
                    # results = cursor.fetchall()
                    # for r in results:
                    #     campo1 = r[0]
                    evidencia = ExamenComlexivoGraduados(graduado_id=int(request.POST['idgraduado']),
                                                         itemexamencomplexivo_id=int(request.POST['idex']),
                                                         examen=null_to_decimal(request.POST['valor'], 2),
                                                         ponderacion=null_to_decimal((null_to_decimal(request.POST['valor'], 2)) / 2, 2)
                                                         # ponderacion=campo1
                                                         )
                    evidencia.save(request)
                    log(u'Adicionó Examen Complexivo graduado: %s' % evidencia, request, "add")
                graduado = Graduado.objects.get(pk=int(request.POST['idgraduado']))
                promedio = graduado.promediotitulacion
                if ExamenComlexivoGraduados.objects.filter(graduado=graduado, status=True).exists():
                    promediotitulacion = ExamenComlexivoGraduados.objects.filter(graduado=graduado, status=True).aggregate(promedio=Avg('examen'))['promedio']
                    graduado.promediotitulacion = null_to_decimal(promediotitulacion, 2)
                    graduado.save(request)
                    log(u'Se actualizo nota del examen complexivo en graduados: valor ingresado (%s), promedio titulacion antes de actualizar(%s), despues de actualizar (%s)' % (request.POST['valor'], promedio.__str__(), graduado.promediotitulacion), request, "add")
                return JsonResponse({'result': 'ok', 'valor': 5})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad'})

        elif action == 'updatefecha':
            try:
                if ExamenComlexivoGraduados.objects.filter(pk=request.POST['mid']).exists():
                    detalle = ExamenComlexivoGraduados.objects.get(pk=request.POST['mid'])
                    detalle.fecha = request.POST['fecha']
                    detalle.save(request)
                    log(u'Editó Examen Complexivo graduado: %s' % detalle, request, "edit")
                return JsonResponse({'result': 'ok', 'valor': 5})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad'})

        elif action == 'pdfacta':
            try:
                data = {}
                data['graduado'] = graduado = Graduado.objects.get(inscripcion=request.POST['id'])
                abrsecretaria = 'S'#'E'
                if CargoInstitucion.objects.filter(persona=graduado.secretariageneral, status=True):
                    nombressecretatia = graduado.secretariageneral.cargoinstitucion_set.get(status=True)
                    abrsecretaria = nombressecretatia.abreviatura
                data['persona'] = persona
                data['abrsecretaria'] = abrsecretaria
                data['promfinal'] = promediofinal = graduado.calcular_notagraduacion()
                # data['promfinal'] = promediofinal = null_to_decimal(((graduado.notafinal + graduado.promediotitulacion) / 2),2)
                data['nombrepromediofinal'] = numero_a_letras(promediofinal)
                data['fechagraduados'] = fecha_letra(str(graduado.fechagraduado))
                if graduado.horagraduacion in [None, '', 0]:
                    hora = '00:00'
                    data['horagraduacion'] = '00:00'
                else:
                    hora = str(graduado.horagraduacion).split(':')
                    data['horagraduacion'] = hora[0] + ':' + hora[1]
                if graduado.horacertificacion in [None, '', 0]:
                    horacer = '00:00'
                    data['horacertificacion'] = '00:00'
                else:
                    horacer = str(graduado.horacertificacion).split(':')
                    data['horacertificacion'] = horacer[0] + ':' + horacer[1]
                data['nomtitulacion'] = numero_a_letras(graduado.promediotitulacion)
                if graduado.fechaconsejo in [None, '']:
                    data['fechaconsejoanio'] = ''
                    data['fechaconsejomes'] = ''
                    data['fechaconsejodia'] = ''
                else:
                    cadenafecha = str(graduado.fechaconsejo).split('-')
                    data['fechaconsejoanio'] = int(cadenafecha[0])
                    data['fechaconsejomes'] = mes[int(cadenafecha[1])]
                    data['fechaconsejodia'] = int(cadenafecha[2])
                return conviert_html_to_pdf(
                    'graduados/actagrado_pdf.html',
                    {
                        'pagesize': 'A4',
                        'data': data,
                    }
                )
            except Exception as ex:
                pass

        elif action == 'pdfactagradpoposgrado':
            try:
                actaposgrado = actagradoposgrado(request.POST['id'])
                return actaposgrado
            except Exception as ex:
                pass

        elif action == 'pdfactagradpoposgradograduado':
            try:
                actaposgrado = getactagradoposgradograduados2(request.POST['id'])
                return actaposgrado
            except Exception as ex:
                pass

        elif action == 'graduadopersona':
            try:
                graduado = Graduado.objects.get(pk=int(request.POST['id']))
                return JsonResponse({"result": "ok", 'estudiante': graduado.inscripcion.persona.nombre_completo_inverso(), 'idaf': graduado.actafacultad.id if graduado.actafacultad else None})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al consultar los datos."})

        elif action == 'addactafacultad':
            try:
                graduado = Graduado.objects.get(pk=int(request.POST['idg']))
                if not request.POST['idaf'] == '':
                    actafacultad = ActaFacultad.objects.get(pk=int(request.POST['idaf']))
                else:
                    actafacultad = None
                graduado.actafacultad = actafacultad
                graduado.save(request)
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'pdfactasustentacion':
            try:
                participante = ComplexivoDetalleGrupo.objects.filter(actatribunalgenerada=True,grupo__activo=True, numeroacta__gt=0, actacerrada=True, matricula__estado=10, matricula__estadotitulacion=3, matricula__inscripcion_id=request.POST['id'], status=True)[0]
                actatribunal = actatribunalcalificacion(participante.id)
                return actatribunal
            except Exception as ex:
                pass

        elif action == 'pdfrubricasustentacion':
            try:
                data = {}
                lista = []
                participante = ComplexivoDetalleGrupo.objects.filter(actatribunalgenerada=True,grupo__activo=True, numeroacta__gt=0, actacerrada=True, matricula__estado=10, matricula__estadotitulacion=3, matricula__inscripcion_id=request.POST['id'], status=True)[0]
                rubricatribunal = rubricatribunalcalificacion(participante.id)
                return rubricatribunal
            except Exception as ex:
                pass

        elif action == 'pdfactaexcomplexivo':
            try:
                data = {}
                data['graduado'] = graduado = Graduado.objects.get(inscripcion=request.POST['id'])
                # graduado.calcular_notagraduacion()
                abrsecretaria = 'S'#'E'
                if CargoInstitucion.objects.filter(persona=graduado.secretariageneral, status=True):
                    nombressecretatia = graduado.secretariageneral.cargoinstitucion_set.get(status=True)
                    abrsecretaria = nombressecretatia.abreviatura
                data['abrsecretaria'] = abrsecretaria
                exados = None
                data['calificacion'] = exauno = None
                data['calificacionletra'] = 0
                notaexcomplexivo = None
                if ExamenComlexivoGraduados.objects.filter(graduado=graduado).exists():
                    if ExamenComlexivoGraduados.objects.filter(graduado=graduado, itemexamencomplexivo_id=1).exists():
                        data['calificacion'] = exauno = ExamenComlexivoGraduados.objects.get(graduado=graduado, itemexamencomplexivo_id=1)
                        data['calificacionletra'] = numero_a_letras(exauno.examen)
                if graduado.matriculatitulacion:
                    exados = None
                    notaexcomplexivo = None
                    if not graduado.matriculatitulacion.alternativa.aplicapropuesta:
                        if ExamenComlexivoGraduados.objects.filter(graduado=graduado, itemexamencomplexivo_id=2).exists():
                            exados = ExamenComlexivoGraduados.objects.get(graduado=graduado, itemexamencomplexivo_id=2)
                            if exauno:
                                notaexcomplexivo = null_to_decimal(((exauno.ponderacion + exados.ponderacion)), 2)
                            else:
                                notaexcomplexivo = 0
                    else:
                        if ExamenComlexivoGraduados.objects.filter(graduado=graduado, itemexamencomplexivo_id=3).exists():
                            exados = ExamenComlexivoGraduados.objects.filter(graduado=graduado, itemexamencomplexivo_id=3)[0]
                            if exauno:
                                notaexcomplexivo = null_to_decimal(((exauno.ponderacion + exados.ponderacion)), 2)
                            else:
                                notaexcomplexivo = 0
                    data['calificaciondos'] = exados
                    data['totalexcomplexivo'] = null_to_decimal((notaexcomplexivo), 2)
                    data['totalexcomplexivoletra'] = numero_a_letras(null_to_decimal((notaexcomplexivo), 2))
                else:
                    if ExamenComlexivoGraduados.objects.filter(graduado=graduado, itemexamencomplexivo_id=2).exists():
                        exados = ExamenComlexivoGraduados.objects.get(graduado=graduado, itemexamencomplexivo_id=2)
                        if exauno:
                            notaexcomplexivo = null_to_decimal(((exauno.ponderacion + exados.ponderacion)), 2)
                        else:
                            notaexcomplexivo = 0
                    else:
                        if ExamenComlexivoGraduados.objects.filter(graduado=graduado, itemexamencomplexivo_id=3).exists():
                            exados = ExamenComlexivoGraduados.objects.filter(graduado=graduado, itemexamencomplexivo_id=3)[0]
                            if exauno:
                                notaexcomplexivo = null_to_decimal(((exauno.ponderacion + exados.ponderacion)), 2)
                            else:
                                notaexcomplexivo = 0
                    data['calificaciondos'] = exados
                    data['totalexcomplexivo'] = null_to_decimal((notaexcomplexivo), 2)
                    data['totalexcomplexivoletra'] = numero_a_letras(null_to_decimal((notaexcomplexivo), 2))

                nomtitulacion = None
                promfinal = None
                nombrepromediofinal = None
                if exados:
                    nomtitulacion = numero_a_letras(notaexcomplexivo)
                promfinal = null_to_decimal((graduado.notagraduacion), 2)
                # promfinal = null_to_decimal(((graduado.notafinal + notaexcomplexivo)/2),2)
                nombrepromediofinal = numero_a_letras(promfinal)
                data['nomtitulacion'] = nomtitulacion
                data['promfinal'] = promfinal
                data['nombrepromediofinal'] = nombrepromediofinal
                data['fechagraduados'] = fecha_letra(str(graduado.fechagraduado)) if graduado.fechagraduado else None
                if graduado.horagraduacion in [None, '', 0]:
                    hora = '00:00'
                    data['horagraduacion'] = '00:00'
                else:
                    hora = str(graduado.horagraduacion).split(':')
                    data['horagraduacion'] = hora[0] + ':' + hora[1]
                if graduado.horacertificacion in [None, '', 0]:
                    horacer = '00:00'
                    data['horacertificacion'] = '00:00'
                else:
                    horacer = str(graduado.horacertificacion).split(':')
                    data['horacertificacion'] = horacer[0] + ':' + horacer[1]
                if graduado.fechaconsejo in [None, '']:
                    data['fechaconsejoanio'] = '____'
                    data['fechaconsejomes'] = '__'
                    data['fechaconsejodia'] = '__'
                else:
                    cadenafecha = str(graduado.fechaconsejo).split('-')
                    data['fechaconsejoanio'] = int(cadenafecha[0])
                    data['fechaconsejomes'] = mes[int(cadenafecha[1])]
                    data['fechaconsejodia'] = int(cadenafecha[2])
                destino = 'graduados/actagradoexcomplexivo_pdf.html'
                if graduado.matriculatitulacion:
                    if graduado.matriculatitulacion.alternativa.aplicapropuesta:
                        destino = 'graduados/actarecordexcompact_pdf.html'

                data['totalexcomplexivo'] = graduado.promediotitulacion
                data['totalexcomplexivoletra'] = numero_a_letras(null_to_decimal((graduado.promediotitulacion), 2))

                return conviert_html_to_pdf(destino,
                                            {
                                                'pagesize': 'A4',
                                                'data': data,
                                            }
                                            )
            except Exception as ex:
                pass

        elif action == 'pdfactaborrador':
            try:
                data = {}
                data['graduado'] = graduado = Graduado.objects.get(inscripcion=request.POST['id'])
                abrsecretaria = 'S'#'E'
                if CargoInstitucion.objects.filter(persona=graduado.secretariageneral, status=True):
                    nombressecretatia = graduado.secretariageneral.cargoinstitucion_set.get(status=True)
                    abrsecretaria = nombressecretatia.abreviatura
                data['abrsecretaria'] = abrsecretaria
                data['promfinal'] = promediofinal = null_to_decimal(((graduado.notafinal + graduado.promediotitulacion) / 2), 2)
                data['nombrepromediofinal'] = numero_a_letras(promediofinal)
                data['fechagraduados'] = fecha_letra(str(graduado.fechagraduado))
                if graduado.horagraduacion in [None, '', 0]:
                    hora = '00:00'
                    data['horagraduacion'] = '00:00'
                else:
                    hora = str(graduado.horagraduacion).split(':')
                    data['horagraduacion'] = hora[0] + ':' + hora[1]
                if graduado.horacertificacion in [None, '', 0]:
                    horacer = '00:00'
                    data['horacertificacion'] = '00:00'
                else:
                    horacer = str(graduado.horacertificacion).split(':')
                    data['horacertificacion'] = horacer[0] + ':' + horacer[1]
                data['nomtitulacion'] = numero_a_letras(graduado.promediotitulacion)
                if graduado.fechaconsejo in [None, '']:
                    data['fechaconsejoanio'] = ''
                    data['fechaconsejomes'] = ''
                    data['fechaconsejodia'] = ''
                else:
                    cadenafecha = str(graduado.fechaconsejo).split('-')
                    data['fechaconsejoanio'] = int(cadenafecha[0])
                    data['fechaconsejomes'] = mes[int(cadenafecha[1])]
                    data['fechaconsejodia'] = int(cadenafecha[2])
                return conviert_html_to_pdf(
                    'graduados/actagradoborrador_pdf.html',
                    {
                        'pagesize': 'A4',
                        'data': data,
                    }
                )
            except Exception as ex:
                pass

        elif action == 'pdfactaborradorexcom':
            try:
                data = {}
                data['graduado'] = graduado = Graduado.objects.get(inscripcion=request.POST['id'])
                abrsecretaria = 'S'#'E'
                if CargoInstitucion.objects.filter(persona=graduado.secretariageneral, status=True):
                    nombressecretatia = graduado.secretariageneral.cargoinstitucion_set.get(status=True)
                    abrsecretaria = nombressecretatia.abreviatura
                data['abrsecretaria'] = abrsecretaria
                data['promfinal'] = promediofinal = null_to_decimal(((graduado.notafinal + graduado.promediotitulacion) / 2), 2)
                data['nombrepromediofinal'] = numero_a_letras(promediofinal)
                data['fechagraduados'] = fecha_letra(str(graduado.fechagraduado))
                if graduado.horagraduacion in [None, '', 0]:
                    hora = '00:00'
                    data['horagraduacion'] = '00:00'
                else:
                    hora = str(graduado.horagraduacion).split(':')
                    data['horagraduacion'] = hora[0] + ':' + hora[1]
                if graduado.horacertificacion in [None, '', 0]:
                    horacer = '00:00'
                    data['horacertificacion'] = '00:00'
                else:
                    horacer = str(graduado.horacertificacion).split(':')
                    data['horacertificacion'] = horacer[0] + ':' + horacer[1]
                data['nomtitulacion'] = numero_a_letras(graduado.promediotitulacion)
                if graduado.fechaconsejo in [None, '']:
                    data['fechaconsejoanio'] = ''
                    data['fechaconsejomes'] = ''
                    data['fechaconsejodia'] = ''
                else:
                    cadenafecha = str(graduado.fechaconsejo).split('-')
                    data['fechaconsejoanio'] = int(cadenafecha[0])
                    data['fechaconsejomes'] = mes[int(cadenafecha[1])]
                    data['fechaconsejodia'] = int(cadenafecha[2])
                return conviert_html_to_pdf(
                    'graduados/actagradoborradorexcom_pdf.html',
                    {
                        'pagesize': 'A4',
                        'data': data,
                    }
                )
            except Exception as ex:
                pass

        elif action == 'pdfactarecord':
            try:
                data = {}
                promediofinal = 0
                data['graduado'] = graduado = Graduado.objects.get(inscripcion=request.POST['id'])
                data['directores'] = graduado.directoresfacultad.all()
                data['es_complexivo'] = False
                abrsecretaria = 'S'#'E'
                if CargoInstitucion.objects.filter(persona=graduado.secretariageneral, status=True):
                    nombressecretatia = graduado.secretariageneral.cargoinstitucion_set.get(status=True)
                    abrsecretaria = nombressecretatia.abreviatura
                data['abrsecretaria'] = abrsecretaria
                if ExamenComlexivoGraduados.objects.filter(graduado=graduado).exists():
                    data['es_complexivo'] = True
                    data['calificacion'] = exauno = None
                    if ExamenComlexivoGraduados.objects.filter(graduado=graduado, itemexamencomplexivo_id=1).exists():
                        data['calificacion'] = exauno = ExamenComlexivoGraduados.objects.get(graduado=graduado, itemexamencomplexivo_id=1)
                    if exauno:
                        if ExamenComlexivoGraduados.objects.filter(graduado=graduado, itemexamencomplexivo_id=2).exists():
                            data['calificaciondos'] = exados = ExamenComlexivoGraduados.objects.get(graduado=graduado, itemexamencomplexivo_id=2)
                            data['totalexcomplexivo'] = notaexcomplexivo = null_to_decimal((exauno.ponderacion + exados.ponderacion), 2)
                            data['nomtitulacion'] = numero_a_letras(notaexcomplexivo)
                            # promediofinal = null_to_decimal(((graduado.notafinal + notaexcomplexivo) / 2),2)
                            promediofinal = graduado.notagraduacion
                        else:
                            if ExamenComlexivoGraduados.objects.filter(graduado=graduado, itemexamencomplexivo_id=3).exists():
                                data['calificaciondos'] = exados = ExamenComlexivoGraduados.objects.get(graduado=graduado, itemexamencomplexivo_id=3)
                                if exauno:
                                    data['totalexcomplexivo'] = notaexcomplexivo = null_to_decimal((exauno.ponderacion + exados.ponderacion), 2)
                                    data['nomtitulacion'] = numero_a_letras(notaexcomplexivo)
                                else:
                                    data['totalexcomplexivo'] = notaexcomplexivo = 0
                                    data['nomtitulacion'] = "NO HAY NOTA"
                                # promediofinal = null_to_decimal(((graduado.notafinal + notaexcomplexivo) / 2), 2)
                                promediofinal = graduado.notagraduacion
                # else:
                # promediofinal = null_to_decimal(((graduado.notafinal + graduado.promediotitulacion) / 2), 2)
                promediofinal = graduado.notagraduacion
                data['totaltitulacion'] = null_to_decimal(graduado.promediotitulacion, 2)

                if graduado.fechagraduado in [None, '']:
                    data['fechagraduados'] = ''
                else:
                    data['fechagraduados'] = fecha_letra(str(graduado.fechagraduado))
                if graduado.fechaconsejo in [None, '']:
                    data['fechaconsejo'] = ''
                else:
                    data['fechaconsejo'] = fecha_letra(str(graduado.fechaconsejo))
                # if PracticasPreprofesionalesInscripcion.objects.filter(inscripcion_id=request.POST['id']).exists():
                #     data['practicasprofesionales'] = PracticasPreprofesionalesInscripcion.objects.filter(inscripcion_id=request.POST['id'])
                if graduado.fechaconsejo in [None, '']:
                    data['mesconsejo'] = ''
                else:
                    data['mesconsejo'] = fecha_letra(str(graduado.fechaconsejo))
                if Egresado.objects.filter(inscripcion=request.POST['id'], status=True).exists():
                    data['egresado'] = Egresado.objects.get(inscripcion=request.POST['id'], status=True)
                if PerfilInscripcion.objects.filter(persona=graduado.inscripcion.persona.id).exists():
                    data['perfilinscripcion'] = PerfilInscripcion.objects.get(persona=graduado.inscripcion.persona.id)
                else:
                    data['perfilinscripcion'] = ''

                # data['horasvinculacion'] = null_to_numeric(ParticipantesMatrices.objects.filter(inscripcion=graduado.inscripcion.id, matrizevidencia_id=2, proyecto__tipo=1, proyecto__status=True, status=True).aggregate(valor=Sum('horas'))['valor'])
                data['listacertificadoidioma'] = CertificadoIdioma.objects.values_list('id', 'idioma__nombre', 'nivelsuficencia__descripcion').filter(persona=graduado.inscripcion.persona,
                                                                                                                                                      historialcertificacionpersona__perfilusuario=graduado.inscripcion.perfil_usuario(),
                                                                                                                                                      historialcertificacionpersona__estado=1,
                                                                                                                                                      status=True).distinct()
                data['horasvinculacion'] = graduado.inscripcion.numero_horas_proyectos_vinculacion()
                data['vinculacion'] = graduado.inscripcion.mis_proyectos_vinculacion()
                # data['vinculacion'] = ParticipantesMatrices.objects.filter(matrizevidencia_id=2, status=True, proyecto__status=True, proyecto__tipo=1, inscripcion_id=graduado.inscripcion.id)
                if PracticasPreprofesionalesInscripcion.objects.filter(inscripcion_id=graduado.inscripcion.id, status=True, culminada=True).exists():
                    data['parcticas'] = PracticasPreprofesionalesInscripcion.objects.filter(inscripcion_id=graduado.inscripcion.id, status=True, culminada=True)
                    data['horaspracticas'] = null_to_numeric(PracticasPreprofesionalesInscripcion.objects.filter(inscripcion=graduado.inscripcion.id, status=True, culminada=True).aggregate(valor=Sum('numerohora'))['valor'])

                data['recordhistorico'] = niveles = RecordAcademico.objects.filter(inscripcion=request.POST['id'],
                                                                                   validapromedio=True,
                                                                                   aprobada=True,
                                                                                   noaplica=False,
                                                                                   asignatura__modulo=False).order_by('asignaturamalla__nivelmalla__id')
                # data['recordhistoricomodulos'] = modingles = RecordAcademico.objects.filter(inscripcion=request.POST['id'],
                #                                                                             aprobada=True,
                #                                                                             noaplica=False,
                #                                                                             asignatura__modulo=True
                #                                                                             ).order_by('asignatura__nombre')
                # data['recordhistoricomodulosingles'] = modingles.filter(asignatura__nombre__contains='INGLES')
                # data['recordhistoricomoduloscomputacion'] = modingles.all().exclude(asignatura__nombre__contains='INGLES')

                asignaturaingles = AsignaturaMalla.objects.values_list('asignatura__id', flat=True).filter(malla__id=22, status=True)
                data['modulo_ingles'] = RecordAcademico.objects.filter(inscripcion__id=request.POST['id'], asignatura__id__in=asignaturaingles, aprobada=True, noaplica=False).order_by('asignatura__nombre')
                asignaturaingles = AsignaturaMalla.objects.values_list('asignatura__id', flat=True).filter(malla__id=32, status=True)
                data['modulo_computacion'] = RecordAcademico.objects.filter(inscripcion__id=request.POST['id'], asignatura__id__in=asignaturaingles, aprobada=True, noaplica=False)
                if graduado.decano:
                    data['decano_es_hombre'] = graduado.decano.es_hombre()
                if graduado.subdecano:
                    data['subdecano_es_hombre'] = graduado.subdecano.es_hombre()
                data['totalhoras'] = graduado.inscripcion.total_horas()
                if graduado.creditotitulacion:
                    data['totalcreditos'] = null_to_decimal(graduado.inscripcion.creditos() + graduado.creditovinculacion + graduado.creditopracticas + graduado.creditotitulacion, 2)
                else:
                    data['totalcreditos'] = null_to_decimal((graduado.inscripcion.creditos() + graduado.creditovinculacion + graduado.creditopracticas + 0), 2)
                data['promfinal'] = promediofinal
                data['asistentefacultad'] = str(graduado.asistentefacultad.nombre_completo()) if graduado.asistentefacultad else None
                return conviert_html_to_pdf(
                    'graduados/actarecord_pdf.html',
                    {
                        'pagesize': 'A4',
                        'data': data,
                    }
                )
            except Exception as ex:
                pass

        elif action == 'pdfactarecord2':
            try:
                data = {}
                promediofinal = 0
                data['graduado'] = graduado = Graduado.objects.get(inscripcion=request.POST['id'])
                data['directores'] = graduado.directoresfacultad.all()
                data['es_complexivo'] = False
                abrsecretaria = 'S'#'E'
                if CargoInstitucion.objects.filter(persona=graduado.secretariageneral, status=True):
                    nombressecretatia = graduado.secretariageneral.cargoinstitucion_set.get(status=True)
                    abrsecretaria = nombressecretatia.abreviatura
                data['abrsecretaria'] = abrsecretaria
                if ExamenComlexivoGraduados.objects.filter(graduado=graduado).exists():
                    data['es_complexivo'] = True
                    data['calificacion'] = exauno = None
                    if ExamenComlexivoGraduados.objects.filter(graduado=graduado, itemexamencomplexivo_id=1).exists():
                        data['calificacion'] = exauno = ExamenComlexivoGraduados.objects.get(graduado=graduado, itemexamencomplexivo_id=1)
                    if exauno:
                        if ExamenComlexivoGraduados.objects.filter(graduado=graduado, itemexamencomplexivo_id=2).exists():
                            data['calificaciondos'] = exados = ExamenComlexivoGraduados.objects.get(graduado=graduado, itemexamencomplexivo_id=2)
                            data['totalexcomplexivo'] = notaexcomplexivo = null_to_decimal((exauno.ponderacion + exados.ponderacion), 2)
                            data['nomtitulacion'] = numero_a_letras(notaexcomplexivo)
                            # promediofinal = null_to_decimal(((graduado.notafinal + notaexcomplexivo) / 2),2)
                            promediofinal = graduado.notagraduacion
                        else:
                            if ExamenComlexivoGraduados.objects.filter(graduado=graduado, itemexamencomplexivo_id=3).exists():
                                data['calificaciondos'] = exados = ExamenComlexivoGraduados.objects.get(graduado=graduado, itemexamencomplexivo_id=3)
                                if exauno:
                                    data['totalexcomplexivo'] = notaexcomplexivo = null_to_decimal((exauno.ponderacion + exados.ponderacion), 2)
                                    data['nomtitulacion'] = numero_a_letras(notaexcomplexivo)
                                else:
                                    data['totalexcomplexivo'] = notaexcomplexivo = 0
                                    data['nomtitulacion'] = "NO HAY NOTA"
                                # promediofinal = null_to_decimal(((graduado.notafinal + notaexcomplexivo) / 2), 2)
                                promediofinal = graduado.notagraduacion
                # else:
                # promediofinal = null_to_decimal(((graduado.notafinal + graduado.promediotitulacion) / 2), 2)
                promediofinal = graduado.notagraduacion
                data['totaltitulacion'] = null_to_decimal(graduado.promediotitulacion, 2)

                if graduado.fechagraduado in [None, '']:
                    data['fechagraduados'] = ''
                else:
                    data['fechagraduados'] = fecha_letra(str(graduado.fechagraduado))
                if graduado.fechaconsejo in [None, '']:
                    data['fechaconsejo'] = ''
                else:
                    data['fechaconsejo'] = fecha_letra(str(graduado.fechaconsejo))
                # if PracticasPreprofesionalesInscripcion.objects.filter(inscripcion_id=request.POST['id']).exists():
                #     data['practicasprofesionales'] = PracticasPreprofesionalesInscripcion.objects.filter(inscripcion_id=request.POST['id'])
                if graduado.fechaconsejo in [None, '']:
                    data['mesconsejo'] = ''
                else:
                    data['mesconsejo'] = fecha_letra(str(graduado.fechaconsejo))
                if Egresado.objects.filter(inscripcion=request.POST['id'], status=True).exists():
                    data['egresado'] = Egresado.objects.get(inscripcion=request.POST['id'], status=True)
                if PerfilInscripcion.objects.filter(persona=graduado.inscripcion.persona.id).exists():
                    data['perfilinscripcion'] = PerfilInscripcion.objects.get(persona=graduado.inscripcion.persona.id)
                else:
                    data['perfilinscripcion'] = ''

                # data['horasvinculacion'] = null_to_numeric(ParticipantesMatrices.objects.filter(inscripcion=graduado.inscripcion.id, matrizevidencia_id=2, proyecto__tipo=1, proyecto__status=True, status=True).aggregate(valor=Sum('horas'))['valor'])
                data['horasvinculacion'] = graduado.inscripcion.numero_horas_proyectos_vinculacion()
                data['vinculacion'] = graduado.inscripcion.mis_proyectos_vinculacion()
                # data['vinculacion'] = ParticipantesMatrices.objects.filter(matrizevidencia_id=2, status=True, proyecto__status=True, proyecto__tipo=1, inscripcion_id=graduado.inscripcion.id)
                if PracticasPreprofesionalesInscripcion.objects.filter(inscripcion_id=graduado.inscripcion.id, status=True, culminada=True).exists():
                    data['parcticas'] = PracticasPreprofesionalesInscripcion.objects.filter(inscripcion_id=graduado.inscripcion.id, status=True, culminada=True)
                    data['horaspracticas'] = null_to_numeric(PracticasPreprofesionalesInscripcion.objects.filter(inscripcion=graduado.inscripcion.id, status=True, culminada=True).aggregate(valor=Sum('numerohora'))['valor'])

                data['recordhistorico'] = niveles = RecordAcademico.objects.filter(inscripcion=request.POST['id'],
                                                                                   validapromedio=True,
                                                                                   aprobada=True,
                                                                                   noaplica=False,
                                                                                   asignatura__modulo=False).order_by('asignaturamalla__nivelmalla__id')
                # data['recordhistoricomodulos'] = modingles = RecordAcademico.objects.filter(inscripcion=request.POST['id'],
                #                                                                             aprobada=True,
                #                                                                             noaplica=False,
                #                                                                             asignatura__modulo=True
                #                                                                             ).order_by('asignatura__nombre')
                # data['recordhistoricomodulosingles'] = modingles.filter(asignatura__nombre__contains='INGLES')
                # data['recordhistoricomoduloscomputacion'] = modingles.all().exclude(asignatura__nombre__contains='INGLES')

                asignaturaingles = AsignaturaMalla.objects.values_list('asignatura__id', flat=True).filter(malla__id=22, status=True)
                data['modulo_ingles'] = RecordAcademico.objects.filter(inscripcion__id=request.POST['id'], asignatura__id__in=asignaturaingles, aprobada=True, noaplica=False).order_by('asignatura__nombre')
                asignaturaingles = AsignaturaMalla.objects.values_list('asignatura__id', flat=True).filter(malla__id=32, status=True)
                data['modulo_computacion'] = RecordAcademico.objects.filter(inscripcion__id=request.POST['id'], asignatura__id__in=asignaturaingles, aprobada=True, noaplica=False)
                if graduado.decano:
                    data['decano_es_hombre'] = graduado.decano.es_hombre()
                if graduado.subdecano:
                    data['subdecano_es_hombre'] = graduado.subdecano.es_hombre()
                data['totalhoras'] = graduado.inscripcion.total_horas()
                if graduado.creditotitulacion:
                    data['totalcreditos'] = null_to_decimal(graduado.inscripcion.creditos() + graduado.creditovinculacion + graduado.creditopracticas + graduado.creditotitulacion, 2)
                else:
                    data['totalcreditos'] = null_to_decimal((graduado.inscripcion.creditos() + graduado.creditovinculacion + graduado.creditopracticas + 0), 2)
                data['promfinal'] = promediofinal
                data['asistentefacultad'] = str(graduado.asistentefacultad.nombre_completo()) if graduado.asistentefacultad else None
                return conviert_html_to_pdf(
                    'graduados/actarecord_pdf2.html',
                    {
                        'pagesize': 'A4',
                        'data': data,
                    }
                )
            except Exception as ex:
                pass

        elif action == 'pdfrecordintegracion':
            try:
                data = {}
                promediofinal = 0
                data['graduado'] = graduado = Graduado.objects.get(inscripcion=request.POST['id'])
                data['directores'] = graduado.directoresfacultad.all()
                data['es_complexivo'] = False
                abrsecretaria = 'S'#'E'
                if CargoInstitucion.objects.filter(persona=graduado.secretariageneral, status=True):
                    nombressecretatia = graduado.secretariageneral.cargoinstitucion_set.get(status=True)
                    abrsecretaria = nombressecretatia.abreviatura
                data['abrsecretaria'] = abrsecretaria
                if ExamenComlexivoGraduados.objects.filter(graduado=graduado).exists():
                    data['es_complexivo'] = True
                    data['calificacion'] = exauno = None
                    if ExamenComlexivoGraduados.objects.filter(graduado=graduado, itemexamencomplexivo_id=1).exists():
                        data['calificacion'] = exauno = ExamenComlexivoGraduados.objects.get(graduado=graduado, itemexamencomplexivo_id=1)
                    if exauno:
                        if ExamenComlexivoGraduados.objects.filter(graduado=graduado, itemexamencomplexivo_id=2).exists():
                            data['calificaciondos'] = exados = ExamenComlexivoGraduados.objects.get(graduado=graduado, itemexamencomplexivo_id=2)
                            data['totalexcomplexivo'] = notaexcomplexivo = null_to_decimal((exauno.ponderacion + exados.ponderacion), 2)
                            data['nomtitulacion'] = numero_a_letras(notaexcomplexivo)
                            # promediofinal = null_to_decimal(((graduado.notafinal + notaexcomplexivo) / 2),2)
                            promediofinal = graduado.notagraduacion
                        else:
                            if ExamenComlexivoGraduados.objects.filter(graduado=graduado, itemexamencomplexivo_id=3).exists():
                                data['calificaciondos'] = exados = ExamenComlexivoGraduados.objects.get(graduado=graduado, itemexamencomplexivo_id=3)
                                if exauno:
                                    data['totalexcomplexivo'] = notaexcomplexivo = null_to_decimal((exauno.ponderacion + exados.ponderacion), 2)
                                    data['nomtitulacion'] = numero_a_letras(notaexcomplexivo)
                                else:
                                    data['totalexcomplexivo'] = notaexcomplexivo = 0
                                    data['nomtitulacion'] = "NO HAY NOTA"
                                # promediofinal = null_to_decimal(((graduado.notafinal + notaexcomplexivo) / 2), 2)
                                promediofinal = graduado.notagraduacion
                # else:
                # promediofinal = null_to_decimal(((graduado.notafinal + graduado.promediotitulacion) / 2), 2)
                promediofinal = graduado.notagraduacion
                data['totaltitulacion'] = null_to_decimal(graduado.promediotitulacion, 2)

                if graduado.fechagraduado in [None, '']:
                    data['fechagraduados'] = ''
                else:
                    data['fechagraduados'] = fecha_letra(str(graduado.fechagraduado))
                if graduado.fechaconsejo in [None, '']:
                    data['fechaconsejo'] = ''
                else:
                    data['fechaconsejo'] = fecha_letra(str(graduado.fechaconsejo))
                # if PracticasPreprofesionalesInscripcion.objects.filter(inscripcion_id=request.POST['id']).exists():
                #     data['practicasprofesionales'] = PracticasPreprofesionalesInscripcion.objects.filter(inscripcion_id=request.POST['id'])
                if graduado.fechaconsejo in [None, '']:
                    data['mesconsejo'] = ''
                else:
                    data['mesconsejo'] = fecha_letra(str(graduado.fechaconsejo))
                if Egresado.objects.filter(inscripcion=request.POST['id'], status=True).exists():
                    data['egresado'] = Egresado.objects.get(inscripcion=request.POST['id'], status=True)
                if PerfilInscripcion.objects.filter(persona=graduado.inscripcion.persona.id).exists():
                    data['perfilinscripcion'] = PerfilInscripcion.objects.get(persona=graduado.inscripcion.persona.id)
                else:
                    data['perfilinscripcion'] = ''

                # data['horasvinculacion'] = null_to_numeric(ParticipantesMatrices.objects.filter(inscripcion=graduado.inscripcion.id, matrizevidencia_id=2, proyecto__tipo=1, proyecto__status=True, status=True).aggregate(valor=Sum('horas'))['valor'])
                data['horasvinculacion'] = graduado.inscripcion.numero_horas_proyectos_vinculacion()
                data['vinculacion'] = graduado.inscripcion.mis_proyectos_vinculacion()
                data['listacertificadoidioma'] = CertificadoIdioma.objects.values_list('id', 'idioma__nombre','nivelsuficencia__descripcion').filter(persona=graduado.inscripcion.persona,
                                                                                                                                                     historialcertificacionpersona__perfilusuario=graduado.inscripcion.perfil_usuario(),
                                                                                                                                                     historialcertificacionpersona__estado=1,
                                                                                                                                                     status=True).distinct()
                # data['vinculacion'] = ParticipantesMatrices.objects.filter(matrizevidencia_id=2, status=True, proyecto__status=True, proyecto__tipo=1, inscripcion_id=graduado.inscripcion.id)
                if PracticasPreprofesionalesInscripcion.objects.filter(inscripcion_id=graduado.inscripcion.id, status=True, culminada=True).exists():
                    data['parcticas'] = PracticasPreprofesionalesInscripcion.objects.filter(inscripcion_id=graduado.inscripcion.id, status=True, culminada=True)
                    data['horaspracticas'] = null_to_numeric(PracticasPreprofesionalesInscripcion.objects.filter(inscripcion=graduado.inscripcion.id, status=True, culminada=True).aggregate(valor=Sum('numerohora'))['valor'])

                data['recordhistorico'] = niveles = RecordAcademico.objects.filter(inscripcion=request.POST['id'],
                                                                                   validapromedio=True,
                                                                                   aprobada=True,
                                                                                   noaplica=False,
                                                                                   asignatura__modulo=False,
                                                                                   status=True).order_by('asignaturamalla__nivelmalla__id')
                # data['recordhistoricomodulos'] = modingles = RecordAcademico.objects.filter(inscripcion=request.POST['id'],
                #                                                                             aprobada=True,
                #                                                                             noaplica=False,
                #                                                                             asignatura__modulo=True
                #                                                                             ).order_by('asignatura__nombre')
                # data['recordhistoricomodulosingles'] = modingles.filter(asignatura__nombre__contains='INGLES')
                # data['recordhistoricomoduloscomputacion'] = modingles.all().exclude(asignatura__nombre__contains='INGLES')

                asignaturaingles = AsignaturaMalla.objects.values_list('asignatura__id', flat=True).filter(malla__id=22, status=True)
                data['modulo_ingles'] = RecordAcademico.objects.filter(inscripcion__id=request.POST['id'], asignatura__id__in=asignaturaingles, aprobada=True, noaplica=False, status=True).order_by('asignatura__nombre')
                asignaturaingles = AsignaturaMalla.objects.values_list('asignatura__id', flat=True).filter(malla__id=32, status=True)
                data['modulo_computacion'] = RecordAcademico.objects.filter(inscripcion__id=request.POST['id'], asignatura__id__in=asignaturaingles, aprobada=True, noaplica=False, status=True)
                if graduado.decano:
                    data['decano_es_hombre'] = graduado.decano.es_hombre()
                if graduado.subdecano:
                    data['subdecano_es_hombre'] = graduado.subdecano.es_hombre()
                data['totalhoras'] = graduado.inscripcion.total_horas()
                if graduado.creditotitulacion:
                    data['totalcreditos'] = null_to_decimal(graduado.inscripcion.creditos() + graduado.creditovinculacion + graduado.creditopracticas + graduado.creditotitulacion, 2)
                else:
                    data['totalcreditos'] = null_to_decimal((graduado.inscripcion.creditos() + graduado.creditovinculacion + graduado.creditopracticas + 0), 2)
                data['promfinal'] = promediofinal
                data['asistentefacultad'] = str(graduado.asistentefacultad.nombre_completo()) if graduado.asistentefacultad else None
                return conviert_html_to_pdf(
                    'graduados/recordintegracion_pdf.html',
                    {
                        'pagesize': 'A4',
                        'data': data,
                    }
                )
            except Exception as ex:
                pass

        elif action == 'pdfactarecordexcomp':
            try:
                data = {}
                data['graduado'] = graduado = Graduado.objects.get(inscripcion=request.POST['id'])
                data['es_complexivo'] = False
                exauno = None
                notaexcomplexivo = 0
                if ExamenComlexivoGraduados.objects.filter(graduado=graduado).exists():
                    if ExamenComlexivoGraduados.objects.filter(graduado=graduado, itemexamencomplexivo_id=1).exists():
                        data['calificacion'] = exauno = ExamenComlexivoGraduados.objects.get(graduado=graduado, itemexamencomplexivo_id=1)
                        data['es_complexivo'] = True
                if ExamenComlexivoGraduados.objects.filter(graduado=graduado, itemexamencomplexivo_id=2).exists():
                    data['calificaciondos'] = exados = ExamenComlexivoGraduados.objects.get(graduado=graduado, itemexamencomplexivo_id=2)
                    if exauno:
                        data['totalexcomplexivo'] = notaexcomplexivo = exauno.ponderacion + exados.ponderacion
                        data['nomtitulacion'] = numero_a_letras(notaexcomplexivo)
                        data['promfinal'] = promediofinal = null_to_decimal(((graduado.notafinal + notaexcomplexivo) / 2), 2)
                    else:
                        data['totalexcomplexivo'] = 0
                        data['nomtitulacion'] = "NO HAY REGISTRO"
                        data['promfinal'] = 0
                if graduado.fechagraduado in [None, '']:
                    data['fechagraduados'] = ''
                else:
                    data['fechagraduados'] = fecha_letra(str(graduado.fechagraduado))
                if graduado.fechaconsejo in [None, '']:
                    data['fechaconsejo'] = ''
                else:
                    data['fechaconsejo'] = fecha_letra(str(graduado.fechaconsejo))
                if PracticasPreprofesionalesInscripcion.objects.filter(inscripcion_id=request.POST['id']).exists():
                    data['practicasprofesionales'] = PracticasPreprofesionalesInscripcion.objects.filter(inscripcion_id=request.POST['id'])
                if graduado.fechaconsejo in [None, '']:
                    data['mesconsejo'] = ''
                else:
                    data['mesconsejo'] = fecha_letra(str(graduado.fechaconsejo))
                if Egresado.objects.filter(inscripcion=request.POST['id'], status=True).exists():
                    data['egresado'] = Egresado.objects.get(inscripcion=request.POST['id'], status=True)
                if PerfilInscripcion.objects.filter(persona=graduado.inscripcion.persona.id).exists():
                    data['perfilinscripcion'] = PerfilInscripcion.objects.get(persona=graduado.inscripcion.persona.id)
                else:
                    data['perfilinscripcion'] = ''
                data['horasvinculacion'] = null_to_numeric(ParticipantesMatrices.objects.filter(inscripcion=graduado.inscripcion.id, matrizevidencia_id=2, proyecto__tipo=1, proyecto__status=True, status=True).aggregate(valor=Sum('horas'))['valor'])
                data['recordhistorico'] = niveles = RecordAcademico.objects.filter(inscripcion=request.POST['id'],
                                                                                   validapromedio=True,
                                                                                   aprobada=True,
                                                                                   noaplica=False,
                                                                                   asignatura__modulo=False).order_by('asignaturamalla__nivelmalla__id')
                data['recordhistoricomodulos'] = modingles = RecordAcademico.objects.filter(inscripcion=request.POST['id'],
                                                                                            aprobada=True,
                                                                                            noaplica=False,
                                                                                            asignatura__modulo=True
                                                                                            ).order_by('asignatura__nombre')
                # data['recordhistoricomodulosingles'] = modingles.filter(asignatura__nombre__contains='INGLES')
                # data['recordhistoricomoduloscomputacion'] = modingles.all().exclude(asignatura__nombre__contains='INGLES')

                asignaturaingles = AsignaturaMalla.objects.values_list('asignatura__id', flat=True).filter(malla__id=22, status=True)
                data['modulo_ingles'] = RecordAcademico.objects.filter(inscripcion__id=request.POST['id'], asignatura__id__in=asignaturaingles, aprobada=True, noaplica=False)
                asignaturaingles = AsignaturaMalla.objects.values_list('asignatura__id', flat=True).filter(malla__id=32, status=True)
                data['modulo_computacion'] = RecordAcademico.objects.filter(inscripcion__id=request.POST['id'], asignatura__id__in=asignaturaingles, aprobada=True, noaplica=False)
                data['decano_es_hombre'] = graduado.decano.es_hombre()
                data['subdecano_es_hombre'] = graduado.subdecano.es_hombre()
                data['totalhoras'] = graduado.inscripcion.total_horas()
                data['totalcreditos'] = graduado.inscripcion.creditos()
                return conviert_html_to_pdf(
                    'graduados/actarecordexcomp_pdf.html',
                    {
                        'pagesize': 'A4',
                        'data': data,
                    }
                )
            except Exception as ex:
                pass

        elif action == 'pdfcalificaciontitulacion':
            try:
                data = {}
                data['graduado'] = graduado = Graduado.objects.get(inscripcion=request.POST['id'])
                if graduado.fechagraduado in [None, '']:
                    data['fechagraduadoanio'] = ''
                    data['fechagraduadomes'] = ''
                    data['fechagraduadodia'] = ''
                else:
                    cadenafecha = str(graduado.fechagraduado).split('-')
                    data['fechagraduadoanio'] = int(cadenafecha[0])
                    data['fechagraduadomes'] = mes[int(cadenafecha[1])]
                    data['fechagraduadodia'] = int(cadenafecha[2])
                return conviert_html_to_pdf(
                    'graduados/calificacion_titulacion_pdf.html',
                    {
                        'pagesize': 'A4',
                        'data': data,
                    }
                )
            except Exception as ex:
                pass

        elif action == 'pdfcalificacionexamencomplexivo':
            try:
                data = {}
                data['graduado'] = graduado = Graduado.objects.get(inscripcion=request.POST['id'])
                exauno = None
                if ExamenComlexivoGraduados.objects.filter(graduado=graduado, itemexamencomplexivo_id=1).exists():
                    data['calificacion'] = exauno = ExamenComlexivoGraduados.objects.get(graduado=graduado, itemexamencomplexivo_id=1)
                total = 0
                calificaciondos = 0
                if exauno:
                    if ExamenComlexivoGraduados.objects.filter(graduado=graduado, itemexamencomplexivo_id=2).exists():
                        calificaciondos = exados = ExamenComlexivoGraduados.objects.get(graduado=graduado, itemexamencomplexivo_id=2)
                        total = null_to_decimal((exauno.ponderacion + exados.ponderacion), 2)
                    # total = null_to_decimal((exados.ponderacion),2)
                if ExamenComlexivoGraduados.objects.filter(graduado=graduado, itemexamencomplexivo_id=3).exists():
                    calificaciondos = exados = ExamenComlexivoGraduados.objects.get(graduado=graduado, itemexamencomplexivo_id=3)
                    if exauno:
                        total = null_to_decimal((exauno.ponderacion + exados.ponderacion), 2)
                    else:
                        total = 0
                    # total = null_to_decimal((exados.ponderacion), 2)
                data['calificaciondos'] = calificaciondos
                data['total'] = total
                if graduado.fechagraduado in [None, '']:
                    data['fechagraduadoanio'] = ''
                    data['fechagraduadomes'] = ''
                    data['fechagraduadodia'] = ''
                else:
                    cadenafecha = str(graduado.fechagraduado).split('-')
                    data['fechagraduadoanio'] = int(cadenafecha[0])
                    data['fechagraduadomes'] = mes[int(cadenafecha[1])]
                    data['fechagraduadodia'] = int(cadenafecha[2])
                return conviert_html_to_pdf(
                    'graduados/calificacion_examencomplex_pdf.html',
                    {
                        'pagesize': 'A4',
                        'data': data,
                    }
                )
            except Exception as ex:
                pass

        elif action == 'del':
            try:
                graduado = Graduado.objects.get(pk=request.POST['id'])
                graduado.delete()
                log(u'Eliminado graduado: %s' % graduado, request, "del")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'delsubdecano':
            try:
                graduado = Graduado.objects.get(pk=request.POST['id'])
                graduado.subdecano = None
                graduado.save(request)
                log(u'Eliminó decano graduado: %s' % graduado, request, "del")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'delrepresentantedocente':
            try:
                graduado = Graduado.objects.get(pk=request.POST['id'])
                graduado.representantedocente = None
                graduado.save(request)
                log(u'Eliminó decano graduado: %s' % graduado, request, "del")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'delrepresentanteestudiantil':
            try:
                graduado = Graduado.objects.get(pk=request.POST['id'])
                graduado.representanteestudiantil = None
                graduado.save(request)
                log(u'Eliminó representante estudiantil en graduado: %s' % graduado, request, "del")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'resetear':
            try:
                inscripcion = Inscripcion.objects.get(pk=request.POST['id'])
                resetear_clave(inscripcion.persona)
                log(u'Reseteo clave de inscripcion: %s' % inscripcion, request, "add")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'editarnumeroacta':
            try:
                graduado = Graduado.objects.get(pk=request.POST['id'])
                graduado.numeroactagrado = pk = request.POST['idnumeroacta']
                graduado.save()
                log(u'edito numero acta: %s' % graduado, request, "edit")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'addactasgraduados':
            try:
                graduado = Graduado.objects.get(pk=int(request.POST['id']))
                newfile = None
                nombre = request.POST['acta']
                newfile = request.FILES['archivo']
                if newfile.size > 10430400:
                    return JsonResponse({"result": True, "mensaje": u"Error, el tamaño del archivo es mayor a 10 Mb."})
                newfile._name = generar_nombre(nombre.replace(" ",""), newfile._name)
                ruta = 'actas/graduados/' + newfile._name


                if graduado:
                    if TipoActa.objects.filter(nombre=nombre, graduado=graduado):
                        tipo = TipoActa.objects.get(nombre=nombre, graduado=graduado)
                        tipo.archivo = newfile
                        tipo.save(request)
                    else:
                        tipo = TipoActa.objects.create(nombre=nombre, graduado=graduado, archivo=newfile)
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                return JsonResponse({"result": "bad"})

        elif action == 'adicionaractagraduado':
            try:
                form = ActaGraduadosForm(request.POST, request.FILES)
                if form.is_valid():
                    if 'archivo' in request.FILES:
                        arch = request.FILES['archivo']
                        extension = arch._name.split('.')
                        tam = len(extension)
                        exte = extension[tam - 1]
                        if arch.size > 20971520:
                            return JsonResponse(
                                {"result": "bad", "mensaje": u"Error, el tamaño del archivo es mayor a 20 Mb."})
                        if not exte.lower() == 'pdf':
                            return JsonResponse({"result": "bad", "mensaje": u"Solo se permiten archivos .pdf "})
                        newfile = request.FILES['archivo']
                        newfile._name = generar_nombre("actagraduado", newfile._name)
                        # verificaracta = TipoActa.objects.filter(graduado_id=int(request.POST['id']), tipo=form.cleaned_data['acta'], status=True)
                        verificaracta = TipoActa.objects.filter(graduado_id=int(request.POST['id']), tipo=request.POST['tipoacta'], status=True)
                        if not verificaracta:
                            tipoactagraduado = TipoActa(graduado_id=int(request.POST['id']), tipo=request.POST['tipoacta'],
                                                        archivo=newfile)
                            tipoactagraduado.save(request)
                        else:
                            verificaracta[0].archivo = newfile
                            verificaracta[0].save(request)
                        return JsonResponse({"result": False})
                    return JsonResponse({"result": True, "mensaje": "El periodo del inventario ha finalizado"})
                else:
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                         "mensaje": "Error en el formulario"})
            except Exception as ex:
                return JsonResponse({"result": True, "mensaje": "Por favor, llene el formulario correctamente"})

        elif action == 'reporte_certificado':
            try:
                graduados = Graduado.objects.filter(pk=int(request.POST['id']))
                if graduados.first().fecharefrendacion:
                    if graduados.first().fecharefrendacion.year >= 2022:
                        # if graduados.first().fecharefrendacion <= datetime.now().date():
                        resultados_errores = generartituloeinsigniaposgrado(request, graduados, data, IS_DEBUG=IS_DEBUG)
                        if len(resultados_errores) > 0:
                            return JsonResponse({"result": "bad", "mensaje": u"Problemas al generar el Título. "+ str(resultados_errores)})
                        else:
                            return JsonResponse({"result": "ok", "mensaje": u"Título generado exitosamente"})
                    else:
                        return JsonResponse({"result": "bad", "mensaje": u"Acción no permitida, porque la fecha de refrendación del graduado no corresponde al año 2022 o superiores."})
                else:
                    return JsonResponse({"result": "bad", "mensaje": u"Debe ingresar la fecha de refrendación del estudiante de posgrado."})
            except Exception as ex:
                print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
                messages.error(request, ex)
                return JsonResponse({"result": "bad", "mensaje": u"Problemas al ejecutar el reporte. %s" % ex})

        elif action == 'notificar_certificado':
            try:
                graduados = Graduado.objects.filter(pk=int(request.POST['id']), inscripcion__carrera__coordinacion__id=7, rutapdftitulo__isnull=False, urlhtmltitulo__isnull=False, namehtmltitulo__isnull=False)
                if graduados.first().fecharefrendacion:
                    if graduados.first().fecharefrendacion.year >= 2022:
                        if graduados:
                            resulnotificacionerrores = solonotificarcorreoinsigniaposgrado(request, graduados, procesomasivo=False, IS_DEBUG=IS_DEBUG)
                            if len(resulnotificacionerrores) > 0:
                                return JsonResponse({"result": "bad", "mensaje": u"Problemas al enviar el email."})
                            else:
                                return JsonResponse({"result": "ok", "mensaje": u"Email enviado correctamente."})
                        else:
                            return JsonResponse({"result": "bad", "mensaje": u"Por favor, genere primero el Título e Insignia."})
                    else:
                        return JsonResponse({"result": "bad", "mensaje": u"Acción no permitida, porque la fecha de refrendación del graduado no corresponde al año 2022 o superiores."})
                else:
                    return JsonResponse({"result": "bad", "mensaje": u"Debe ingresar la fecha de refrendación del estudiante de posgrado."})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Problemas al enviar la notificación al email. %s" % ex})

        elif action == 'reporte_certificadoprevio':
            try:
                graduados = Graduado.objects.filter(pk=int(request.POST['id']))
                resulfuncion = generartituloeinsigniaposgrado(request, graduados, data, vistaprevia = True, IS_DEBUG=IS_DEBUG)
                # if resulfuncion['result'] == 'ok':
                #     return JsonResponse({"result": "ok"})
                # else:
                #     return JsonResponse({"result": "bad", "mensaje": u"Problemas al ejecutar el reporte."})
                return resulfuncion
            except Exception as ex:
                print(str(ex))
                messages.error(request, ex)

        elif action == 'pdfactatitulacioncomplexivo':
            try:
                idasignadotitulacion = request.POST['idmateriaasign']
                asignado = MateriaTitulacion.objects.get(pk=idasignadotitulacion)
                if asignado.actatitulacionfirmada:
                    # qrname = 'qr_actatitulacion_' + str(asignado.id)
                    # actatribunal = 'https://sga.unemi.edu.ec//media/qrcode/actatitulacion/' + qrname + '_firmado.pdf'
                    actatribunal = 'https://sga.unemi.edu.ec//media/' + str(asignado.archivotitulacionfirmada)
                    # actatribunal = 'http://127.0.0.1:8000/media/qrcode/actatitulacion/' + qrname + '_firmado.pdf'
                else:
                    actatribunal = actatitulacioncomplexivo(idasignadotitulacion)
                    if actatribunal:
                        qrname = 'qr_actatitulacion_' + str(asignado.id)
                        actatribunal = 'https://sga.unemi.edu.ec//media/qrcode/actatitulacion/' + qrname + '.pdf'
                        # actatribunal = 'http://127.0.0.1:8000/media/qrcode/actatitulacion/' + qrname + '.pdf'
                return JsonResponse({"result": "ok", 'url': actatribunal})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'pdfactagradocomplexivo':
            try:
                idgraduado = request.POST['idgraduado']
                actatribunal = actagradocomplexivo(idgraduado)
                return actatribunal
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'importarrefrendacionpos':
            try:
                with transaction.atomic():
                    if not 'archivo_excel' in request.FILES:
                        raise NameError('Carge un archivo excel para ejecutar la acción.')
                    graduados = Graduado.objects.filter(status=True, inscripcion__carrera__coordinacion__id=7)
                    actualizados = 0
                    excel = request.FILES['archivo_excel']
                    wb = openpyxl.load_workbook(excel)
                    lista = wb.worksheets[0]
                    totallista = lista.rows
                    a = 0
                    c = 0
                    for filas in totallista:
                        a += 1
                        if a > 1:
                            id = str(filas[0].value).strip() if filas[0].value else None
                            if not id:
                                raise NameError(f'Error: En la fila {a} la columna ID está vacía.')
                            try:
                                id = int(id)
                            except:
                                raise NameError(f'Error: El formato de la columna ID de la fila {a} es incorrecto. Por favor ingrese un número en la celda.')
                            # if not isinstance(int(id), int):
                            #     raise NameError(f'Error: En la fila {a} la columna ID tiene formato incorrecto. Por favor debe escribir un número.'
                            graduado = graduados.get(pk=id)
                            if graduado:
                                fechacelda = str(filas[12].value).strip()
                                if len(fechacelda) > 0:
                                    if len(fechacelda) > 10:
                                        fechacelda = fechacelda[:len(fechacelda) - 9]
                                    if '/' in fechacelda:
                                        fecha = datetime.strptime(fechacelda, '%d/%m/%Y').date()
                                    else:
                                        if len(str(fechacelda).split('-').__getitem__(0)) == 4:
                                            fecha = datetime.strptime(fechacelda, '%Y-%m-%d').date()
                                        else:
                                            fecha = datetime.strptime(fechacelda, '%d-%m-%Y').date()
                                    if fecha:
                                        # graduado.fecharefrendacion = date(int(fecha[0:4]), int(fecha[5:7]), int(fecha[8:10]))
                                        graduado.fecharefrendacion = date(fecha.year, fecha.month, fecha.day)
                                        graduado.save()
                                        actualizados = actualizados + 1
                                        log(u'Importó fecha de refrendación posgrado %s' % (graduado), request, "edit")
                                    else:
                                        raise NameError(f'Error: En la fila {a} el formato de fecha es incorrecto. Por favor respetar el formato de fecha del documento de excel(formato).')
                                else:
                                    raise NameError(f'Error: En la fila {a} la columna FECHA REFRENDACIÓN está vacía.')
                            else:
                                raise NameError(f'Error: El graduado de código {id} no existe o no pertenece a posgrado.')
                    messages.success(request, f'Importación de fecha de refrendación posgrado realizada correctamente. Total actualizados: {str(actualizados)}.')
                    return JsonResponse({"result": False, "mensaje": f"Importación de fecha de refrendación posgrado realizada correctamente. Total actualizados: {str(actualizados)}."}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                # messages.error(request, u'%s. Línea %s'%(str(ex), sys.exc_info()[-1].tb_lineno))
                return JsonResponse({"result": True, "mensaje": u'%s. (line %s)'%(str(ex), sys.exc_info()[-1].tb_lineno)})

        elif action == 'firmarinformemasivo':
            try:
                import json
                informesselect = request.POST['ids'].split(',')
                bandera = False
                p12 = None
                listainscripcion = []
                nombresmae = ''
                conterrornombre = 0
                conteoerror = 0

                firma = request.FILES["firma"]
                passfirma = request.POST['palabraclave']
                txtFirmas = json.loads(request.POST['txtFirmas'])
                razon = request.POST['razon'] if 'razon' in request.POST else ''
                bytes_certificado = firma.read()
                if not txtFirmas:
                    return JsonResponse({"result": "bad", "mensaje": u"No se ha podido seleccionar la ubicación de la firma"})
                for idinforme in informesselect:
                    try:
                        siguiente = 0
                        historial = TipoActaFirma.objects.get(pk=idinforme)
                        if historial.orden == 1:
                            pdf = historial.tipoacta.archivo
                            pdfname = SITE_STORAGE + '/media/' + str(historial.tipoacta.archivo)
                        else:
                            ultimafirma = TipoActaFirma.objects.filter(tipoacta=historial.tipoacta, firmado=True).order_by('-orden')[0]
                            pdf = ultimafirma.archivofirmado
                            pdfname = SITE_STORAGE + '/media/' + str(ultimafirma.archivofirmado)


                        palabras = 'firma' + str(historial.orden)
                        documento = fitz.open(pdfname)
                        numpaginafirma = int(documento.page_count) - 1
                        with fitz.open(pdfname) as document:
                            words_dict = {}
                            for page_number, page in enumerate(document):
                                if page_number == numpaginafirma:
                                    words = page.get_text("blocks")
                                    words_dict[0] = words
                        valor = None
                        for cadena in words_dict[0]:
                            if palabras in cadena[4]:
                                valor = cadena

                        if valor:
                            posicinony = 5000 - int(valor[3]) - 4125
                        else:
                            messages.warning(request, "Alerta: El nombre en la firma no es el correcto. Se ha rechazado y enviado a comercialización.")
                            return JsonResponse({"result": "errornombre"})
                        x, y, numpaginafirma = obtener_posicion_x_y_saltolinea(pdf.url, palabras)

                        if x and y:
                            x = x - 30
                            y = posicinony

                            extension_certificado = os.path.splitext(firma.name)[1][1:]

                            datau = JavaFirmaEc(
                                archivo_a_firmar=pdf, archivo_certificado=bytes_certificado, extension_certificado=extension_certificado,
                                password_certificado=passfirma,
                                page=int(numpaginafirma), reason=razon, lx=x, ly=y
                            ).sign_and_get_content_bytes()
                            if datau:
                                generar_archivo_firmado = io.BytesIO()
                                generar_archivo_firmado.write(datau)
                                generar_archivo_firmado.seek(0)
                                extension = pdf.name.split('.')
                                tam = len(extension)
                                exte = extension[tam - 1]
                                siguiente = historial.orden + 1
                                if not TipoActaFirma.objects.filter(tipoacta=historial.tipoacta, orden=siguiente):
                                    _name = 'fe_actagrado_firmada_' + str(historial.tipoacta.graduado.id)
                                else:
                                    _name = 'fe_actagrado_' + str(historial.tipoacta.graduado.id) + '_firma' + str(historial.orden)
                                file_obj = DjangoFile(generar_archivo_firmado, name=f"{remover_caracteres_especiales_unicode(_name)}.pdf")

                                archivoexiste = SITE_STORAGE + '/media/actasgraduados/' + _name + '.pdf'
                                if os.path.isfile(archivoexiste):
                                    os.remove(archivoexiste)

                                historial.archivofirmado = file_obj
                                historial.fechafirma = datetime.now().date()
                                historial.firmado = True
                                historial.turnofirmar = False
                                historial.save(request)

                                if TipoActaFirma.objects.filter(tipoacta=historial.tipoacta, orden=siguiente):
                                    siguientefirmar = TipoActaFirma.objects.filter(tipoacta=historial.tipoacta, orden=siguiente)[0]
                                    siguientefirmar.turnofirmar=True
                                    siguientefirmar.save(request)

                                    itemtipoacta = TipoActa.objects.get(pk=historial.tipoacta.id)
                                    itemtipoacta.archivo = 'actasgraduados/' + _name + '.pdf'
                                    itemtipoacta.save(request)
                                else:
                                    itemtipoacta = TipoActa.objects.get(pk=historial.tipoacta.id)
                                    itemtipoacta.actafirmada=True
                                    itemtipoacta.archivo='actasgraduados/' + _name + '.pdf'
                                    itemtipoacta.save(request)
                    except Exception as ex:
                        pass
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                eline = 'Error on line {} {}'.format(sys.exc_info()[-1].tb_lineno, ex.__str__())
                return JsonResponse({"result": "bad", "mensaje": u"Error: %s" % eline})

        elif action == 'firmaractasconsolidadasmasivo':
            try:
                import json
                informesselect = request.POST['ids'].split(',')
                firma = request.FILES["firma"]
                passfirma = request.POST['palabraclave']
                txtFirmas = json.loads(request.POST['txtFirmas'])
                razon = request.POST['razon'] if 'razon' in request.POST else ''
                bytes_certificado = firma.read()
                for idinforme in informesselect:
                    try:
                        siguiente = 0
                        historial = TipoActaFirma.objects.get(pk=idinforme)
                        if historial.orden == 1:
                            pdf = historial.tipoacta.archivo
                            pdfname = SITE_STORAGE + '/media/' + str(historial.tipoacta.archivo)
                        else:
                            ultimafirma = TipoActaFirma.objects.filter(tipoacta=historial.tipoacta, firmado=True).order_by('-orden')[0]
                            pdf = ultimafirma.archivofirmado
                            pdfname = SITE_STORAGE + '/media/' + str(ultimafirma.archivofirmado)
                        palabras = 'firma' + str(historial.orden)
                        x, y, numpaginafirma = obtener_posicion_x_y_saltolinea_firmapagina(pdf.url, palabras)
                        if x and y:
                            x = x - 30
                            y = y
                            if not txtFirmas:
                                return JsonResponse({"result": "bad", "mensaje": u"No se ha podido seleccionar la ubicación de la firma"})
                            extension_certificado = os.path.splitext(firma.name)[1][1:]

                            datau = JavaFirmaEc(
                                archivo_a_firmar=pdf, archivo_certificado=bytes_certificado, extension_certificado=extension_certificado,
                                password_certificado=passfirma,
                                page=int(numpaginafirma), reason=razon, lx=x, ly=y
                            ).sign_and_get_content_bytes()
                            if datau:
                                generar_archivo_firmado = io.BytesIO()
                                generar_archivo_firmado.write(datau)
                                generar_archivo_firmado.seek(0)
                                extension = pdf.name.split('.')
                                tam = len(extension)
                                exte = extension[tam - 1]
                                siguiente = historial.orden + 1
                                if not TipoActaFirma.objects.filter(tipoacta=historial.tipoacta, orden=siguiente):
                                    _name = 'fe_actaconsolidada_firmada_' + str(historial.tipoacta.graduado.id)
                                else:
                                    _name = 'fe_actaconsolidada_' + str(historial.tipoacta.graduado.id) + '_firma' + str(historial.orden)
                                file_obj = DjangoFile(generar_archivo_firmado, name=f"{remover_caracteres_especiales_unicode(_name)}.pdf")

                                archivoexiste = SITE_STORAGE + '/media/actasgraduados/' + _name + '.pdf'
                                if os.path.isfile(archivoexiste):
                                    os.remove(archivoexiste)

                                historial.archivofirmado = file_obj
                                historial.fechafirma = datetime.now().date()
                                historial.firmado = True
                                historial.turnofirmar = False
                                historial.save(request)

                                if TipoActaFirma.objects.filter(tipoacta=historial.tipoacta, orden=siguiente):
                                    siguientefirmar = TipoActaFirma.objects.filter(tipoacta=historial.tipoacta, orden=siguiente)[0]
                                    siguientefirmar.turnofirmar=True
                                    siguientefirmar.save(request)

                                    itemtipoacta = TipoActa.objects.get(pk=historial.tipoacta.id)
                                    itemtipoacta.archivo = 'actasgraduados/' + _name + '.pdf'
                                    itemtipoacta.save(request)
                                else:
                                    itemtipoacta = TipoActa.objects.get(pk=historial.tipoacta.id)
                                    itemtipoacta.actafirmada=True
                                    itemtipoacta.archivo='actasgraduados/' + _name + '.pdf'
                                    itemtipoacta.save(request)
                    except Exception as ex:
                        pass
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                eline = 'Error on line {} {}'.format(sys.exc_info()[-1].tb_lineno, ex.__str__())
                return JsonResponse({"result": "bad", "mensaje": u"Error: %s" % eline})

        elif action == 'eliminadetallefirmasactagrado':
            try:
                tipoacta = TipoActa.objects.get(pk=int(request.POST['id']))
                listadofirmas = tipoacta.tipoactafirma_set.filter(status=True).order_by('orden')
                for lisacta in listadofirmas:
                    lisacta.fechafirma = None
                    if lisacta.orden == 1:
                        lisacta.turnofirmar = True
                        lisacta.firmado = False
                    else:
                        lisacta.turnofirmar = False
                        lisacta.firmado = False
                    lisacta.save(request)
                    if lisacta.archivofirmado:
                        archivoexiste = SITE_STORAGE + lisacta.archivofirmado.url
                        if os.path.isfile(archivoexiste):
                            os.remove(archivoexiste)
                nombre1 = 'fe_actagrado_'+ str(tipoacta.graduado.id) +'_firma1'
                archivoexiste1 = SITE_STORAGE + '/media/actasgraduados/' + nombre1 + '.pdf'
                if os.path.isfile(archivoexiste1):
                    os.remove(archivoexiste1)

                nombre2 = 'fe_actagrado_'+ str(tipoacta.graduado.id) +'_firma2'
                archivoexiste2 = SITE_STORAGE + '/media/actasgraduados/' + nombre2 + '.pdf'
                if os.path.isfile(archivoexiste2):
                    os.remove(archivoexiste2)

                nombre3 = 'fe_actagrado_'+ str(tipoacta.graduado.id) +'_firma3'
                archivoexiste3 = SITE_STORAGE + '/media/actasgraduados/' + nombre3 + '.pdf'
                if os.path.isfile(archivoexiste3):
                    os.remove(archivoexiste3)

                nombrefirmada = 'fe_actagrado_firmada_' + str(tipoacta.graduado.id)
                archivofirmada = SITE_STORAGE + '/media/actasgraduados/' + nombrefirmada + '.pdf'
                if os.path.isfile(archivofirmada):
                    os.remove(archivofirmada)

                # Actualizamos horas de la malla al graduado en caso que exista cambios en su malla con las horas
                mallaalumno = tipoacta.graduado.inscripcion.mi_malla()
                if tipoacta.graduado.inscripcion.persona.sexo:
                    if tipoacta.graduado.inscripcion.persona.sexo.id == 1:
                        nombretitulo = mallaalumno.tituloobtenidomujer
                    else:
                        nombretitulo = mallaalumno.tituloobtenidohombre
                    tipoacta.graduado.nombretitulo = nombretitulo
                tipoacta.graduado.horastitulacion = mallaalumno.horas_titulacion
                tipoacta.graduado.creditotitulacion = mallaalumno.creditos_titulacion
                tipoacta.graduado.creditovinculacion = mallaalumno.creditos_vinculacion
                tipoacta.graduado.creditopracticas = mallaalumno.creditos_practicas
                tipoacta.graduado.save(request)

                # generamos otra vez el acta
                generaactagradofe = actagradocomplexivofirma(tipoacta.graduado.id)
                if generaactagradofe:
                    tipoacupdate = TipoActa.objects.filter(tipo=5, graduado=tipoacta.graduado)[0]
                    tipoacupdate.archivo = 'qrcode/actatitulacion/fe_actagrado_' + str(tipoacta.graduado.id) + '.pdf'
                    tipoacupdate.save(request)

                log(u'Reseteo firmas actas grados: %s' % tipoacta, request, "add")
                return JsonResponse({"result": 'ok',"mensaje":'El acta ha sido eliminada!'}, safe=False)
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
                return JsonResponse(res_json, safe=False)

        elif action == 'eliminadetallefirmasactaconsolidada':
            try:
                tipoacta = TipoActa.objects.get(pk=int(request.POST['id']))
                listadofirmas = tipoacta.tipoactafirma_set.filter(status=True).order_by('orden')
                for lisacta in listadofirmas:
                    lisacta.fechafirma = None
                    if lisacta.orden == 1:
                        lisacta.turnofirmar = True
                        lisacta.firmado = False
                    else:
                        lisacta.turnofirmar = False
                        lisacta.firmado = False
                    lisacta.save(request)
                    if lisacta.archivofirmado:
                        archivoexiste = SITE_STORAGE + lisacta.archivofirmado.url
                        if os.path.isfile(archivoexiste):
                            os.remove(archivoexiste)
                nombre1 = 'fe_actaconsolidada_'+ str(tipoacta.graduado.id) +'_firma1'
                archivoexiste1 = SITE_STORAGE + '/media/actasgraduados/' + nombre1 + '.pdf'
                if os.path.isfile(archivoexiste1):
                    os.remove(archivoexiste1)

                nombre2 = 'fe_actaconsolidada_'+ str(tipoacta.graduado.id) +'_firma2'
                archivoexiste2 = SITE_STORAGE + '/media/actasgraduados/' + nombre2 + '.pdf'
                if os.path.isfile(archivoexiste2):
                    os.remove(archivoexiste2)

                nombre3 = 'fe_actaconsolidada_'+ str(tipoacta.graduado.id) +'_firma3'
                archivoexiste3 = SITE_STORAGE + '/media/actasgraduados/' + nombre3 + '.pdf'
                if os.path.isfile(archivoexiste3):
                    os.remove(archivoexiste3)

                nombrefirmada = 'fe_actaconsolidada_firmada_' + str(tipoacta.graduado.id)
                archivofirmada = SITE_STORAGE + '/media/actasgraduados/' + nombrefirmada + '.pdf'
                if os.path.isfile(archivofirmada):
                    os.remove(archivofirmada)

                # Actualizamos horas de la malla al graduado en caso que exista cambios en su malla con las horas
                mallaalumno = tipoacta.graduado.inscripcion.mi_malla()
                if tipoacta.graduado.inscripcion.persona.sexo:
                    if tipoacta.graduado.inscripcion.persona.sexo.id == 1:
                        nombretitulo = mallaalumno.tituloobtenidomujer
                    else:
                        nombretitulo = mallaalumno.tituloobtenidohombre
                    tipoacta.graduado.nombretitulo = nombretitulo
                tipoacta.graduado.horastitulacion = mallaalumno.horas_titulacion
                tipoacta.graduado.creditotitulacion = mallaalumno.creditos_titulacion
                tipoacta.graduado.creditovinculacion = mallaalumno.creditos_vinculacion
                tipoacta.graduado.creditopracticas = mallaalumno.creditos_practicas
                tipoacta.graduado.save(request)
                # generamos otra vez el acta
                generaactaconsolidadafe = actaconsolidadafirma(tipoacta.graduado.id)
                if generaactaconsolidadafe:
                    tipoacupdate = TipoActa.objects.filter(tipo=6, graduado=tipoacta.graduado)[0]
                    tipoacupdate.archivo = 'qrcode/actatitulacion/fe_actaconsolidada_' + str(tipoacta.graduado.id) + '.pdf'
                    tipoacupdate.save(request)
                    # actualiza usuarios que firman
                    listadopersonafirman = tipoacupdate.tipoactafirma_set.filter(status=True)
                    for lpersonas in listadopersonafirman:
                        if lpersonas.orden == 2:
                            for dirfac in tipoacta.graduado.directoresfacultad.all():
                                dirfacultad = dirfac
                            lpersonas.persona = dirfacultad
                            lpersonas.save(request)
                        if lpersonas.orden == 3:
                            lpersonas.persona_id = tipoacta.graduado.decano_id
                            lpersonas.save(request)
                        if lpersonas.orden == 4:
                            lpersonas.persona_id = tipoacta.graduado.secretariageneral_id
                            lpersonas.save(request)

                log(u'Reseteo firmas actas consolidadas: %s' % tipoacta, request, "add")
                return JsonResponse({"result": 'ok',"mensaje":'El acta ha sido eliminada!'}, safe=False)
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
                return JsonResponse(res_json, safe=False)

        return JsonResponse({"result": "bad", "mensaje": u"Solicitud Incorrecta."})
    else:
        if 'action' in request.GET:
            data['action'] = action = request.GET['action']

            if action == 'add':
                try:
                    data['title'] = u'Graduar estudiante'
                    inscripcion = Inscripcion.objects.filter(carrera__in=miscarreras).get(pk=request.GET['id'])
                    data['inscripcion'] = inscripcion
                    pasantias = inscripcion.promedio_pasantias()
                    vcc = inscripcion.promedio_vcc()
                    ppp = inscripcion.promedio_ppp()
                    pmg = inscripcion.promedio_materias_grado()
                    ne = inscripcion.datos_egresado().notaegreso

                    decano = inscripcion.coordinacion.responsable()
                    decano = Profesor.objects.get(persona=decano.persona)
                    reprefacu = inscripcion.coordinacion.representantesfacultad_set.filter(status=True).order_by('-id')[0]

                    if inscripcion.persona.sexo.id == 2:
                        titulo = inscripcion.malla_inscripcion().malla.tituloobtenidohombre
                    else:
                        titulo = inscripcion.malla_inscripcion().malla.tituloobtenidomujer

                    data['iddecano'] = decano
                    data['idrepresentanteestudiantil'] = reprefacu.representanteestudiantil
                    data['idrepresentantedocente'] = Profesor.objects.get(persona=reprefacu.representantedocente)
                    data['idrepresentanteservidores'] = reprefacu.representanteservidores

                    data['idasistentefacultad'] = persona
                    data['idsecretariageneral'] = idsecretariageneral = CargoInstitucion.objects.get(pk=1)

                    form = GraduadoForm(initial={'fechagraduado': datetime.now().date(),
                                                 'notafinal': ne,
                                                 # 'promediogrado': pmg,
                                                 'fechainicio': Inscripcion.objects.get(pk=int(request.GET['id'])).fechainicioprimernivel if Inscripcion.objects.filter(pk=int(request.GET['id'])).exists() else None,
                                                 'fechaegresado': Egresado.objects.get(inscripcion_id=int(request.GET['id'])).fechaegreso if Egresado.objects.filter(inscripcion_id=int(request.GET['id'])).exists() else None,
                                                 'asistentefacultad': persona.id,
                                                 'secretariageneral': idsecretariageneral.persona.id,
                                                 'promediotitulacion': null_to_decimal((pmg / 3), 2),
                                                 'notagraduacion': 0,
                                                 'decano': decano,
                                                 'nombretitulo': titulo
                                                 # 'representanteestudiantil': reprefacultad.representanteestudiantil.persona.id,
                                                 # 'representantedocente': reprefacultad.representantedocente,
                                                 # 'representanteservidores': reprefacultad.representanteservidores
                                                 })
                    form.tiene_propuesta()
                    data['form'] = form
                    return render(request, "graduados/add.html", data)
                except Exception as ex:
                    pass

            elif action == 'edit':
                try:
                    data['title'] = u'Editar estudiante graduado'

                    data['graduado'] = periodo = request.session['periodo']

                    graduado = Graduado.objects.filter(inscripcion__carrera__in=miscarreras).get(pk=request.GET['id'])
                    decano = ResponsableCoordinacion.objects.filter(coordinacion=graduado.inscripcion.carrera.coordinacion_carrera(), periodo=periodo, tipo=1).first()
                    subdecano = ResponsableCoordinacion.objects.filter(coordinacion=graduado.inscripcion.carrera.coordinacion_carrera(), periodo=periodo, tipo=2).first()
                    data['graduado'] = graduado
                    if graduado.profesor_id in [None, '', 0]:
                        data['idprofesor'] = idprofesor = 0
                    else:
                        data['idprofesor'] = idprofesor = graduado.profesor.id
                    if graduado.integrantetribunal_id in [None, '', 0]:
                        data['idintegrantetribunal'] = idintegrantetribunal = 0
                    else:
                        data['idintegrantetribunal'] = idintegrantetribunal = graduado.integrantetribunal.id
                    if graduado.docentesecretario_id in [None, '', 0]:
                        data['iddocentesecretario'] = iddocentesecretario = 0
                    else:
                        data['iddocentesecretario'] = iddocentesecretario = graduado.docentesecretario.id
                    if graduado.asistentefacultad_id in [None, '', 0]:
                        data['idasistentefacultad'] = codigoasistentefacultad = persona
                    else:
                        data['idasistentefacultad'] = codigoasistentefacultad = Persona.objects.get(pk=graduado.asistentefacultad_id)
                    if graduado.secretariageneral:
                        data['idsecretariageneral'] = secretarigeneral = graduado.secretariageneral
                    else:
                        idsecretariageneral = CargoInstitucion.objects.get(pk=1)
                        data['idsecretariageneral'] = secretarigeneral = idsecretariageneral.persona

                    # data['idsecretariageneral'] = secretarigeneral = idsecretariageneral.persona.id

                    if graduado.decano_id in [None, '', 0]:
                        data['iddecano'] = iddecano = decano.persona.profesor().id
                        decano = decano.persona.profesor()
                    else:
                        data['iddecano'] = iddecano = graduado.decano.profesor().id
                        decano = graduado.decano.profesor()
                    if graduado.subdecano_id in [None, '', 0]:
                        if subdecano not in [None]:
                            data['idsubdecano'] = idsubdecano = subdecano.persona.profesor().id
                            subdecano = subdecano.persona.profesor()
                        else:
                            data['idsubdecano'] = idsubdecano = 0
                            subdecano = 0
                    else:
                        data['idsubdecano'] = idsubdecano = graduado.subdecano.profesor().id
                        subdecano = graduado.subdecano.profesor()

                    if graduado.representanteestudiantil_id in [None, '', 0]:
                        data['idrepresentanteestudiantil'] = idrepresentanteestudiantil = 0
                    else:
                        data['idrepresentanteestudiantil'] = idrepresentanteestudiantil = graduado.representanteestudiantil.id

                    if graduado.representanteestudiantildos_id in [None, '', 0]:
                        data['idrepresentanteestudiantildos'] = idrepresentanteestudiantildos = 0
                    else:
                        data['idrepresentanteestudiantildos'] = idrepresentanteestudiantildos = graduado.representanteestudiantildos.id

                    if graduado.representantesuplenteestudiantil_id in [None, '', 0]:
                        data['idrepresentantesuplenteestudiantil'] = idrepresentantesuplenteestudiantil = 0
                    else:
                        data['idrepresentantesuplenteestudiantil'] = idrepresentantesuplenteestudiantil = graduado.representantesuplenteestudiantil.id

                    if graduado.representantesuplenteestudiantildos_id in [None, '', 0]:
                        data['idrepresentantesuplenteestudiantildos'] = idrepresentantesuplenteestudiantildos = 0
                    else:
                        data['idrepresentantesuplenteestudiantildos'] = idrepresentantesuplenteestudiantildos = graduado.representantesuplenteestudiantildos.id

                    if graduado.representantedocente_id in [None, '', 0]:
                        data['idrepresentantedocente'] = idrepresentantedocente = 0
                    else:
                        data['idrepresentantedocente'] = idrepresentantedocente = graduado.representantedocente.profesor().id

                    if graduado.representantedocentedos_id in [None, '', 0]:
                        data['idrepresentantedocentedos'] = idrepresentantedocentedos = 0
                    else:
                        data['idrepresentantedocentedos'] = idrepresentantedocentedos = graduado.representantedocentedos.profesor().id

                    if graduado.representantesuplentedocente_id in [None, '', 0]:
                        data['idrepresentantesuplentedocente'] = idrepresentantesuplentedocente = 0
                    else:
                        data['idrepresentantesuplentedocente'] = idrepresentantesuplentedocente = graduado.representantesuplentedocente.profesor().id

                    if graduado.representantesuplentedocentedos_id in [None, '', 0]:
                        data['idrepresentantesuplentedocentedos'] = idrepresentantesuplentedocentedos = 0
                    else:
                        data['idrepresentantesuplentedocentedos'] = idrepresentantesuplentedocentedos = graduado.representantesuplentedocentedos.profesor().id

                    if graduado.representanteservidores_id in [None, '', 0]:
                        data['idrepresentanteservidores'] = idrepresentanteservidores = 0
                    else:
                        data['idrepresentanteservidores'] = idrepresentanteservidores = graduado.representanteservidores.administrativo().id

                    if graduado.representanteservidoresdos_id in [None, '', 0]:
                        data['idrepresentanteservidoresdos'] = idrepresentanteservidoresdos = 0
                    else:
                        data['idrepresentanteservidoresdos'] = idrepresentanteservidoresdos = graduado.representanteservidoresdos.administrativo().id

                    if graduado.representantesuplenteservidores_id in [None, '', 0]:
                        data['idrepresentantesuplenteservidores'] = idrepresentantesuplenteservidores = 0
                    else:
                        data['idrepresentantesuplenteservidores'] = idrepresentantesuplenteservidores = graduado.representantesuplenteservidores.administrativo().id

                    if graduado.representantesuplenteservidoresdos_id in [None, '', 0]:
                        data['idrepresentantesuplenteservidoresdos'] = idrepresentantesuplenteservidoresdos = 0
                    else:
                        data['idrepresentantesuplenteservidoresdos'] = idrepresentantesuplenteservidoresdos = graduado.representantesuplenteservidoresdos.administrativo().id

                    graduado.calcular_notagraduacion()
                    # if ExamenComlexivoGraduados.objects.filter(graduado=graduado,status=True, graduado__estadograduado=False ).exists():
                    #     promediotitulacion = ExamenComlexivoGraduados.objects.filter(graduado=graduado,status=True).aggregate(promedio=Avg('examen'))['promedio']
                    #     # cursor = connections['sga_select'].cursor()
                    #     # sql = "select null_to_decimal((" + str(promediotitulacion) + "),2)"
                    #     # cursor.execute(sql)
                    #     # results = cursor.fetchall()
                    #     # for r in results:
                    #     #     campo1 = r[0]
                    #     graduado.promediotitulacion = null_to_decimal(promediotitulacion,2)
                    #     # graduado.promediotitulacion = campo1
                    #     graduado.save(request)
                    # else:
                    #     promediotitulacion = graduado.promediotitulacion
                    carrera = graduado.inscripcion.carrera.id
                    malla = graduado.inscripcion.mi_malla()
                    directorcarrera = graduado.inscripcion.carrera.get_director(periodo)
                    if directorcarrera in [None]:
                        data['iddirector'] = director_id = 0
                        director = 0
                    else:
                        data['iddirector'] = director_id = directorcarrera.persona.id
                        director = directorcarrera.persona
                    if graduado.nombretitulo:
                        nombretitulo = graduado.nombretitulo
                    else:
                        if graduado.inscripcion.persona.sexo == 1:
                            nombretitulo = malla.tituloobtenidomujer
                        else:
                            nombretitulo = malla.tituloobtenidohombre
                    notienepropuesta = False
                    if graduado.matriculatitulacion:
                        if graduado.matriculatitulacion.alternativa.aplicapropuesta:
                            notienepropuesta = True
                    form = GraduadoForm()
                    form.initial={'decano': iddecano,
                                  'sudecano': idsubdecano,
                                  'representanteestudiantil': idrepresentanteestudiantil,
                                  'representanteestudiantildos': idrepresentanteestudiantildos,
                                  'representantesuplenteestudiantil': idrepresentantesuplenteestudiantil,
                                  'representantesuplenteestudiantildos': idrepresentantesuplenteestudiantildos,
                                  'representantedocente': idrepresentantedocente,
                                  'representantedocentedos': idrepresentantedocentedos,
                                  'representantesuplentedocente': idrepresentantesuplentedocente,
                                  'representantesuplentedocentedos': idrepresentantesuplentedocentedos,
                                  'representanteservidores': idrepresentanteservidores,
                                  'representanteservidoresdos': idrepresentanteservidoresdos,
                                  'representantesuplenteservidores': idrepresentantesuplenteservidores,
                                  'representantesuplenteservidoresdos': idrepresentantesuplenteservidoresdos,
                                  'profesor': idprofesor,
                                  'integrantetribunal': idintegrantetribunal,
                                  'docentesecretario': iddocentesecretario,
                                  'asistentefacultad': codigoasistentefacultad.id,
                                  'secretariageneral': secretarigeneral.id,
                                  'fechainicio': Inscripcion.objects.get(pk=graduado.inscripcion.id).fechainicioprimernivel if Inscripcion.objects.filter(pk=graduado.inscripcion.id).exists() else None,
                                  'fechaegresado': Egresado.objects.get(inscripcion_id=graduado.inscripcion.id).fechaegreso if Egresado.objects.filter(inscripcion_id=graduado.inscripcion.id).exists() else None,
                                  'fechagraduado': graduado.fechagraduado,
                                  'horagraduacion': str(graduado.horagraduacion),
                                  'horacertificacion': str(graduado.horacertificacion),
                                  'notafinal': graduado.inscripcion.promedio_record(),
                                  'promediotitulacion': graduado.promediotitulacion,
                                  'tematesis': graduado.tematesis,
                                  'registro': graduado.registro,
                                  'numeroactagrado': graduado.numeroactagrado,
                                  'fechaactagrado': graduado.fechaactagrado,
                                  'fecharefrendacion': graduado.fecharefrendacion,
                                  'fechaconsejo': graduado.fechaconsejo,
                                  # 'mecanismotitulacion': graduado.mecanismotitulacion,
                                  'codigomecanismotitulacion': graduado.codigomecanismotitulacion,
                                  'folio': graduado.folio,
                                  'horastitulacion': graduado.horastitulacion,
                                  'creditotitulacion': graduado.creditotitulacion,
                                  'creditopracticas': graduado.creditopracticas,
                                  'nombretitulo': nombretitulo,
                                  'creditovinculacion': graduado.creditovinculacion,
                                  'estadograduado': graduado.estadograduado,
                                  'docenteevaluador1': graduado.docenteevaluador1 if notienepropuesta else 0,
                                  'docenteevaluador2': graduado.docenteevaluador2 if notienepropuesta else 0,
                                  'directorcarrera': graduado.directorcarrera if notienepropuesta else 0,
                                  'notagraduacion': graduado.notagraduacion if graduado.notagraduacion else graduado.calcular_notagraduacion(),
                                  }

                    if graduado.directoresfacultad.exists():
                        form.fields['directoresfacultad'].initial = graduado.directoresfacultad.all()
                    else:
                        form.fields['directoresfacultad'].initial = director_id
                    # form.fields['facultad'].queryset = Coordinacion.objects.filter(status=True)

                    data['iddoeva1'] = 0
                    data['iddoeva2'] = 0
                    if notienepropuesta:
                        form.no_tiene_propuesta()
                        if graduado.docenteevaluador1:
                            form.fields['docenteevaluador1'].widget.attrs['descripcion'] = graduado.docenteevaluador1
                            form.fields['docenteevaluador1'].widget.attrs['value'] = graduado.docenteevaluador1.id
                            data['iddoeva1'] = graduado.docenteevaluador1.id
                        else:
                            form.fields['docenteevaluador1'].widget.attrs['descripcion'] = '----------------'
                            form.fields['docenteevaluador1'].widget.attrs['value'] = 0
                        if graduado.docenteevaluador2:
                            form.fields['docenteevaluador2'].widget.attrs['descripcion'] = graduado.docenteevaluador2
                            form.fields['docenteevaluador2'].widget.attrs['value'] = graduado.docenteevaluador2.id
                            data['iddoeva2'] = graduado.docenteevaluador2.id
                        else:
                            form.fields['docenteevaluador2'].widget.attrs['descripcion'] = '----------------'
                            form.fields['docenteevaluador2'].widget.attrs['value'] = 0
                        if graduado.directorcarrera:
                            form.fields['directorcarrera'].widget.attrs['descripcion'] = graduado.directorcarrera
                            form.fields['directorcarrera'].widget.attrs['value'] = graduado.directorcarrera.id
                            data['iddirector'] = graduado.directorcarrera.profesor().id
                        else:
                            form.fields['directorcarrera'].widget.attrs['descripcion'] = '----------------'
                            form.fields['directorcarrera'].widget.attrs['value'] = 0
                    else:
                        form.tiene_propuesta()
                    data['decano'] = decano
                    data['subdecano'] = subdecano
                    data['director'] = director
                    data['form'] = form
                    return render(request, "graduados/edit.html", data)
                except Exception as ex:
                    pass

            elif action == 'editintegracioncurricular':
                try:
                    data['title'] = u'Editar estudiante graduado'
                    data['graduado'] = periodo = request.session['periodo']
                    graduado = Graduado.objects.filter(inscripcion__carrera__in=miscarreras).get(pk=request.GET['id'])

                    if graduado.materiatitulacion:
                        mallaalumno = graduado.inscripcion.mi_malla()
                        if graduado.inscripcion.persona.sexo:
                            if graduado.inscripcion.persona.sexo.id == 1:
                                nombretitulo = mallaalumno.tituloobtenidomujer
                            else:
                                nombretitulo = mallaalumno.tituloobtenidohombre
                            graduado.nombretitulo = nombretitulo
                        graduado.horastitulacion = mallaalumno.horas_titulacion
                        graduado.creditotitulacion = mallaalumno.creditos_titulacion
                        graduado.creditovinculacion = mallaalumno.creditos_vinculacion
                        graduado.creditopracticas = mallaalumno.creditos_practicas
                        graduado.save(request)

                    data['graduado'] = graduado
                    if graduado.profesor_id in [None, '', 0]:
                        data['idprofesor'] = idprofesor = 0
                    else:
                        data['idprofesor'] = idprofesor = graduado.profesor.id

                    if graduado.asistentefacultad_id in [None, '', 0]:
                        data['idasistentefacultad'] = codigoasistentefacultad = persona
                    else:
                        data['idasistentefacultad'] = codigoasistentefacultad = Persona.objects.get(pk=graduado.asistentefacultad_id)
                    if graduado.secretariageneral:
                        data['idsecretariageneral'] = secretarigeneral = graduado.secretariageneral
                    else:
                        idsecretariageneral = CargoInstitucion.objects.get(pk=1)
                        data['idsecretariageneral'] = secretarigeneral = idsecretariageneral.persona

                    if graduado.decano_id in [None, '', 0]:
                        data['iddecano'] = iddecano = 0
                    else:
                        data['iddecano'] = iddecano = graduado.decano.profesor().id
                    graduado.calcular_notagraduacion()
                    notienepropuesta = False
                    if graduado.matriculatitulacion:
                        if graduado.matriculatitulacion.alternativa.aplicapropuesta:
                            notienepropuesta = True
                    form = GraduadoIntegracionCurricularForm(initial={'decano': iddecano,
                                                                      # 'profesor': idprofesor,
                                                                      'asistentefacultad': codigoasistentefacultad.id,
                                                                      'secretariageneral': secretarigeneral.id,
                                                                      'fechainicio': Inscripcion.objects.get(pk=graduado.inscripcion.id).fechainicioprimernivel if Inscripcion.objects.filter(pk=graduado.inscripcion.id).exists() else None,
                                                                      'fechaegresado': Egresado.objects.get(inscripcion_id=graduado.inscripcion.id).fechaegreso if Egresado.objects.filter(inscripcion_id=graduado.inscripcion.id).exists() else None,
                                                                      'fechagraduado': graduado.fechagraduado,
                                                                      'horagraduacion': str(graduado.horagraduacion),
                                                                      'horacertificacion': str(graduado.horacertificacion),
                                                                      'notafinal': graduado.inscripcion.promedio_record(),
                                                                      'promediotitulacion': graduado.promediotitulacion,
                                                                      'tematesis': graduado.tematesis,
                                                                      'registro': graduado.registro,
                                                                      'numeroactagrado': graduado.numeroactagrado,
                                                                      'fechaactagrado': graduado.fechaactagrado,
                                                                      'fecharefrendacion': graduado.fecharefrendacion,
                                                                      'fechaconsejo': graduado.fechaconsejo,
                                                                      # 'mecanismotitulacion': graduado.mecanismotitulacion,
                                                                      'codigomecanismotitulacion': graduado.codigomecanismotitulacion,
                                                                      'folio': graduado.folio,
                                                                      'folioactaconsolidada': graduado.folioactaconsolidada,
                                                                      'numeroactaconsolidada': graduado.numeroactaconsolidada,
                                                                      'horastitulacion': graduado.horastitulacion,
                                                                      'creditotitulacion': graduado.creditotitulacion,
                                                                      'creditopracticas': graduado.creditopracticas,
                                                                      'nombretitulo': graduado.nombretitulo,
                                                                      'creditovinculacion': graduado.creditovinculacion,
                                                                      'estadograduado': graduado.estadograduado,
                                                                      'directorcarrera': graduado.directorcarrera if notienepropuesta else 0,
                                                                      'notagraduacion': graduado.notagraduacion if graduado.notagraduacion else graduado.calcular_notagraduacion(),
                                                                      })
                    if graduado.directoresfacultad.exists():
                        form.fields['directoresfacultad'].initial = graduado.directoresfacultad.all()
                    data['iddirector'] = 0
                    data['form'] = form
                    return render(request, "graduados/editintegracioncurricular.html", data)
                except Exception as ex:
                    pass

            elif action == 'del':
                try:
                    data['title'] = u'Borrar graduación'
                    data['graduado'] = Graduado.objects.filter(inscripcion__carrera__in=miscarreras).get(pk=request.GET['id'])
                    return render(request, "graduados/del.html", data)
                except Exception as ex:
                    pass

            elif action == 'examencomplexivo':
                try:
                    data['title'] = u'Calificacion Examen Complexivo'
                    data['graduado'] = graduado = Graduado.objects.get(pk=request.GET['id'])
                    # matricula = graduado.inscripcion.matriculatitulacion_set.filter(status=True, estado=10)[0]
                    # data['notaexamen'] = notaex = matricula.complexivoexamendetalle_set.filter(status=True, estado=3)[0]
                    iditem = graduado.examencomlexivograduados_set.values_list('itemexamencomplexivo__id', flat=False).filter(status=True)
                    data['examen'] = ItemExamenComplexivo.objects.filter(pk__in=iditem)
                    # data['examen'] = ItemExamenComplexivo.objects.all()
                    return render(request, "graduados/examencomplexivo.html", data)
                except Exception as ex:
                    pass

            elif action == 'reportegraduados':
                try:
                    __author__ = 'Unemi'
                    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
                    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                    title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    style1 = easyxf(num_format_str='D-MMM-YY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('graduados_hoja')
                    ws.write_merge(0, 0, 0, 6, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=graduados' + random.randint(1, 10000).__str__() + '.xls'
                    columns = [
                        (u"FACULTAD", 6000),
                        (u"CARRERA", 6000),
                        (u"CEDULA", 6000),
                        (u"ALUMNO", 6000),
                        (u"FECHA GRADUACION", 6000),
                        (u"NOTA FINAL", 6000),
                        (u"PAIS", 6000),
                        (u"PROVINCIA", 6000),
                        (u"CANTON", 6000),
                        (u"PARROQUIA", 6000),
                        (u"CALLE PRINCIPAL", 6000),
                        (u"CALLE SECUNDARIA", 6000),
                        (u"EMAIL", 6000),
                        (u"EMAIL INSTITUCIONAL", 6000),
                        (u"TELEFONO", 6000),
                        (u"TELEFONO CONVENCIONAL", 6000),
                        (u"SEXO", 6000),
                        (u"LGTBI", 6000),
                        (u"INICIO PRIMER NIVEL", 6000),
                        (u"INICIO CONVALIDACION", 600),
                        (u"COLEGIO", 600)
                    ]
                    row_num = 3
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    cursor = connections['sga_select'].cursor()
                    date_format = xlwt.XFStyle()
                    date_format.num_format_str = 'yyyy/mm/dd'
                    sql = "select co.nombre as Facultad, case ca.mencion when '' then ca.nombre else (ca.nombre||' CON MENCIÓN EN '||ca.mencion) end as Carrera, (p.apellido1||' '||p.apellido2||' '||p.nombres) as alumno, gr.fechagraduado, gr.notafinal, p.cedula, " \
                          "(select nombre from sga_pais where id=p.pais_id) as pais, (select nombre from sga_provincia where id=p.provincia_id) as provincia, (select nombre from sga_canton where id=p.canton_id) as canton, " \
                          "(select nombre from sga_parroquia where id=p.parroquia_id) as parroquia,p.direccion,p.direccion2,p.email,p.emailinst,p.telefono,p.telefono_conv, " \
                          "(select nombre from sga_sexo sexo where sexo.id=p.sexo_id ) as sexo,p.lgtbi,i.fechainicioprimernivel,i.fechainicioconvalidacion, i.colegio as colegio" \
                          " from sga_graduado gr, sga_inscripcion i, sga_persona p, sga_carrera ca, sga_coordinacion_carrera cc, sga_coordinacion co where gr.inscripcion_id=i.id " \
                          " and co.id not in(7) and i.persona_id=p.id and i.carrera_id=ca.id and cc.carrera_id=ca.id and co.id=cc.coordinacion_id order by co.nombre,ca.nombre,gr.fechagraduado"
                    cursor.execute(sql)
                    results = cursor.fetchall()
                    row_num = 4
                    for r in results:
                        i = 0
                        campo1 = r[0]
                        campo2 = r[1]
                        campo3 = r[2]
                        campo4 = r[3]
                        campo5 = r[4]
                        campo6 = r[5]
                        campo7 = r[6]
                        campo8 = r[7]
                        campo9 = r[8]
                        campo10 = r[9]
                        campo11 = r[10]
                        campo12 = r[11]
                        campo13 = r[12]
                        campo14 = r[13]
                        campo15 = r[14]
                        campo16 = r[15]
                        campo17 = r[16]
                        if r[17]:
                            campo18 = 'SI'
                        else:
                            campo18 = 'NO'
                        campo19 = r[18]
                        campo20 = r[19]
                        campo21 = r[20]
                        ws.write(row_num, 0, campo1, font_style2)
                        ws.write(row_num, 1, campo2, font_style2)
                        ws.write(row_num, 2, campo6, font_style2)
                        ws.write(row_num, 3, campo3, font_style2)
                        ws.write(row_num, 4, campo4, style1)
                        ws.write(row_num, 5, campo5, font_style2)
                        ws.write(row_num, 6, campo7, font_style2)
                        ws.write(row_num, 7, campo8, font_style2)
                        ws.write(row_num, 8, campo9, font_style2)
                        ws.write(row_num, 9, campo10, font_style2)
                        ws.write(row_num, 10, campo11, font_style2)
                        ws.write(row_num, 11, campo12, font_style2)
                        ws.write(row_num, 12, campo13, font_style2)
                        ws.write(row_num, 13, campo14, font_style2)
                        ws.write(row_num, 14, campo15, font_style2)
                        ws.write(row_num, 15, campo16, font_style2)
                        ws.write(row_num, 16, campo17, font_style2)
                        ws.write(row_num, 17, campo18, font_style2)
                        ws.write(row_num, 18, campo19, date_format)
                        ws.write(row_num, 19, campo20, date_format)
                        ws.write(row_num, 20, campo21, font_style2)
                        row_num += 1
                    wb.save(response)
                    connection.close()
                    return response
                except Exception as ex:
                    pass

            elif action == 'excelgraduadosposgrado':
                try:
                    idcarrera = int(request.GET['id'])
                    __author__ = 'Unemi'
                    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
                    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                    title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    style1 = easyxf(num_format_str='D-MMM-YY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('graduadosposgrado_hoja')
                    ws.write_merge(0, 0, 0, 6, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=graduadosposgrado' + random.randint(1, 10000).__str__() + '.xls'
                    columns = [
                        (u"ID", 2000),
                        (u"COORDINACIÓN", 7000),
                        (u"CODIGO CARRERA SENESCYT", 4000),
                        (u"CARRERA", 8000),
                        (u"TIPO IDENTIFICACIÓN", 5000),
                        (u"CEDULA", 4000),
                        (u"ALUMNO", 10000),
                        (u"NOMBRES", 6000),
                        (u"APELLIDOS", 6000),
                        (u"TITULO", 6000),
                        (u"FECHA EGRESADO", 5000),
                        (u"FECHA GRADUACIÓN", 5000),
                        (u"FECHA REFRENDACIÓN", 6000),
                        (u"NOTA EGRESO", 4000),
                        (u"NOTA TRAB. TITULACION", 3000),
                        (u"NOTA DE GRADO", 3000),
                        (u"PAIS", 4000),
                        (u"PROVINCIA", 6000),
                        (u"CANTON", 6000),
                        (u"PARROQUIA", 5000),
                        (u"CALLE PRINCIPAL", 8000),
                        (u"CALLE SECUNDARIA", 8000),
                        (u"EMAIL", 7000),
                        (u"EMAIL INSTITUCIONAL", 7000),
                        (u"TELEFONO", 5000),
                        (u"TELEFONO CONVENCIONAL", 6000),
                        (u"SEXO", 3000),
                        (u"LGTBI", 2000),
                        (u"INICIO PRIMER NIVEL", 5000),
                        (u"INICIO CONVALIDACION", 3000),
                        (u"MODALIDAD", 4000)
                    ]
                    row_num = 3
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    cursor = connections['sga_select'].cursor()
                    date_format = xlwt.XFStyle()
                    date_format.num_format_str = 'yyyy/mm/dd'
                    if not idcarrera:
                        sql = """
                                SELECT 
                                        gr.id as Codigo,
                                        co.nombre as coordinacion, 
                                        ca.codigo as codigo_carrera,
                                        ca.nombre as carrera,
                                        p.id,
                                        (p.apellido1||' '||p.apellido2||' '||p.nombres) as alumno, 
                                        p.nombres nombres, (p.apellido1||' '||p.apellido2) apellidos,
                                        ca.titulootorga AS titulo_obtener,
                                        eg.fechaegreso fecha_egreso,
                                        gr.fechagraduado, gr.fecharefrendacion,
                                        gr.notafinal,gr.promediotitulacion, gr.notagraduacion,
                                        pa.nombre as pais, 
                                        pro.nombre as provincia, 
                                        can.nombre as canton, 
                                        par.nombre as parroquia,
                                        p.direccion,p.direccion2,p.email,p.emailinst,p.telefono,p.telefono_conv, 
                                        sex.nombre as sexo,
                                        case WHEN p.lgtbi THEN 'SI' ELSE 'NO' END lgtbi,
                                        i.fechainicioprimernivel,i.fechainicioconvalidacion, 
                                        moda.nombre modalidad                                        
                                FROM
                                    sga_graduado gr
                                    INNER JOIN sga_inscripcion i ON i.id= gr.inscripcion_id
                                    INNER JOIN sga_persona p ON p.id=i.persona_id
                                    INNER JOIN sga_carrera ca ON ca.id=i.carrera_id
                                    LEFT  JOIN sga_coordinacion_carrera cc ON cc.carrera_id=ca.id
                                    LEFT JOIN  sga_coordinacion  co ON co.id=cc.coordinacion_id
                                    LEFT JOIN sga_sexo sex ON sex.id=p.sexo_id
                                    LEFT JOIN sga_pais pa ON pa.id=p.pais_id
                                    LEFT JOIN sga_provincia pro ON pro.id=p.provincia_id
                                    LEFT JOIN sga_canton can ON can.id=p.canton_id
                                    LEFT JOIN sga_parroquia par ON par.id=p.canton_id
                                    RIGHT JOIN sga_inscripcionmalla im ON im."inscripcion_id" = i."id"
                                    LEFT JOIN sga_malla malla ON malla."id" = im."malla_id"
                                    LEFT JOIN sga_egresado eg ON eg.inscripcion_id=i.id
                                    INNER JOIN sga_modalidad moda ON moda.id=i.modalidad_id
                                WHERE 
                                    co.id IN(7)                                
                                order by co.nombre,ca.nombre,gr.fechagraduado;
                            """
                    else:
                        sql = """
                                SELECT 
                                        gr.id as Codigo,
                                        co.nombre as coordinacion, 
                                        ca.codigo as codigo_carrera,
                                        ca.nombre as carrera,
                                        p.id,
                                        (p.apellido1||' '||p.apellido2||' '||p.nombres) as alumno, 
                                        p.nombres nombres, (p.apellido1||' '||p.apellido2) apellidos,
                                        ca.titulootorga AS titulo_obtener,
                                        eg.fechaegreso fecha_egreso,
                                        gr.fechagraduado, gr.fecharefrendacion,
                                        gr.notafinal,gr.promediotitulacion, gr.notagraduacion,
                                        pa.nombre as pais, 
                                        pro.nombre as provincia, 
                                        can.nombre as canton, 
                                        par.nombre as parroquia,
                                        p.direccion,p.direccion2,p.email,p.emailinst,p.telefono,p.telefono_conv, 
                                        sex.nombre as sexo,
                                        case WHEN p.lgtbi THEN 'SI' ELSE 'NO' END lgtbi,
                                        i.fechainicioprimernivel,i.fechainicioconvalidacion, 
                                        moda.nombre modalidad                                        
                                FROM
                                    sga_graduado gr
                                    INNER JOIN sga_inscripcion i ON i.id= gr.inscripcion_id
                                    INNER JOIN sga_persona p ON p.id=i.persona_id
                                    INNER JOIN sga_carrera ca ON ca.id=i.carrera_id
                                    LEFT  JOIN sga_coordinacion_carrera cc ON cc.carrera_id=ca.id
                                    LEFT JOIN  sga_coordinacion  co ON co.id=cc.coordinacion_id
                                    LEFT JOIN sga_sexo sex ON sex.id=p.sexo_id
                                    LEFT JOIN sga_pais pa ON pa.id=p.pais_id
                                    LEFT JOIN sga_provincia pro ON pro.id=p.provincia_id
                                    LEFT JOIN sga_canton can ON can.id=p.canton_id
                                    LEFT JOIN sga_parroquia par ON par.id=p.canton_id
                                    RIGHT JOIN sga_inscripcionmalla im ON im."inscripcion_id" = i."id"
                                    LEFT JOIN sga_malla malla ON malla."id" = im."malla_id"
                                    LEFT JOIN sga_egresado eg ON eg.inscripcion_id=i.id
                                    INNER JOIN sga_modalidad moda ON moda.id=i.modalidad_id
                                WHERE 
                                    co.id IN(7) and ca.id IN(%s)                       
                                order by co.nombre,ca.nombre,gr.fechagraduado;
                            """% (idcarrera)
                    cursor.execute(sql)
                    results = cursor.fetchall()
                    row_num = 4
                    for r in results:
                        i = 0
                        persona = Persona.objects.get(pk=r[4])
                        campo0 = r[0]
                        campo1 = r[1]
                        campo2 = r[2]
                        campo3 = r[3]
                        campo4 = persona.tipo_identificacion_completo()
                        campo5 = persona.identificacion()
                        campo6 = r[5]
                        campo7 = r[6]
                        campo8 = r[7]
                        campo9 = r[8]
                        campo10 = r[9] if r[9] else ''
                        campo11 = r[10] if r[10] else ''
                        campo12 = r[11] if r[11] else ''
                        campo13 = r[12]
                        campo14 = r[13]
                        campo15 = r[14]
                        campo16 = r[15]
                        campo17 = r[16]
                        campo18 = r[17]
                        campo19 = r[18]
                        campo20 = r[19]
                        campo21 = r[20]
                        campo22 = r[21]
                        campo23 = r[22]
                        campo24 = r[23]
                        campo25 = r[24]
                        campo26 = r[25]
                        campo27 = r[26]
                        campo28 = r[27] if r[27] else ''
                        campo29 = r[28] if r[28] else ''
                        campo30 = r[29]

                        ws.write(row_num, 0, campo0, font_style2)
                        ws.write(row_num, 1, campo1, font_style2)
                        ws.write(row_num, 2, campo2, font_style2)
                        ws.write(row_num, 3, campo3, font_style2)
                        ws.write(row_num, 4, campo4, font_style2)
                        ws.write(row_num, 5, campo5, font_style2)
                        ws.write(row_num, 6, campo6, font_style2)
                        ws.write(row_num, 7, campo7, font_style2)
                        ws.write(row_num, 8, campo8, font_style2)
                        ws.write(row_num, 9, campo9, font_style2)
                        ws.write(row_num, 10, campo10, date_format)
                        ws.write(row_num, 11, campo11, date_format)
                        ws.write(row_num, 12, campo12, date_format)
                        ws.write(row_num, 13, campo13, font_style2)
                        ws.write(row_num, 14, campo14, font_style2)
                        ws.write(row_num, 15, campo15, font_style2)
                        ws.write(row_num, 16, campo16, font_style2)
                        ws.write(row_num, 17, campo17, font_style2)
                        ws.write(row_num, 18, campo18, font_style2)
                        ws.write(row_num, 19, campo19, font_style2)
                        ws.write(row_num, 20, campo20, font_style2)
                        ws.write(row_num, 21, campo21, font_style2)
                        ws.write(row_num, 22, campo22, font_style2)
                        ws.write(row_num, 23, campo23, font_style2)
                        ws.write(row_num, 24, campo24, font_style2)
                        ws.write(row_num, 25, campo25, font_style2)
                        ws.write(row_num, 26, campo26, font_style2)
                        ws.write(row_num, 27, campo27, font_style2)
                        ws.write(row_num, 28, campo28, date_format)
                        ws.write(row_num, 29, campo29, date_format)
                        ws.write(row_num, 30, campo30, font_style2)
                        row_num += 1
                    wb.save(response)
                    connection.close()
                    return response
                except Exception as ex:
                    messages.error(request, 'Error al generar el reporte excel.' + str(ex))

            elif action == 'reportegraduados2':
                try:
                    notifi = Notificacion(cuerpo=f'Generación de reporte de graduados',
                                          titulo='Reporte de graduados en proceso',
                                          destinatario=persona,
                                          url='',
                                          prioridad=1, app_label='SGA',
                                          fecha_hora_visible=datetime.now() + timedelta(days=1), tipo=2,
                                          en_proceso=True)
                    notifi.save(request)
                    export_reportegraduados2(request=request, notiid=notifi.id).start()
                    return JsonResponse({"result": True,"mensaje": u"El reporte de graduados se está realizando. Verifique su apartado de notificaciones después de unos minutos.", "btn_notificaciones": traerNotificaciones(request, data, persona)})
                except Exception as ex:
                    pass

            elif action == 'buscarpersona':
                try:
                    q = request.GET['q'].upper().strip()
                    s = q.split(" ")
                    per = Persona.objects.filter(Q(perfilusuario__administrativo__isnull=False) | Q(
                        perfilusuario__profesor__isnull=False)).distinct()

                    if len(s) == 1:
                        per = per.filter(
                            (Q(nombres__icontains=q) | Q(apellido1__icontains=q) | Q(cedula__icontains=q) | Q(
                                apellido2__icontains=q) | Q(cedula__contains=q)),
                            Q(status=True)).distinct()[:15]
                    elif len(s) == 2:
                        per = per.filter(
                            (Q(apellido1__contains=s[0]) & Q(apellido2__contains=s[1])) | (
                                    Q(nombres__icontains=s[0]) & Q(
                                nombres__icontains=s[1])) | (
                                    Q(nombres__icontains=s[0]) & Q(
                                apellido1__contains=s[1]))).filter(status=True).distinct()[
                              :15]
                    else:
                        per = per.filter(
                            (Q(nombres__contains=s[0]) & Q(apellido1__contains=s[1]) & Q(
                                apellido2__contains=s[2])) | (Q(nombres__contains=s[0]) & Q(
                                nombres__contains=s[1]) & Q(apellido1__contains=s[2]))).filter(
                            status=True).distinct()[:15]

                    data = {"result": "ok",
                            "results": [{"id": x.id, "name": "{} - {}".format(x.cedula, x.nombre_completo())}
                                        for x in per]}
                    return JsonResponse(data)
                except Exception as ex:
                    pass


            elif action == 'resetear':
                try:
                    data['title'] = u'Resetear clave del usuario'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    return render(request, "graduados/resetear.html", data)
                except Exception as ex:
                    pass

            elif action == 'excelgraduados':
                try:
                    __author__ = 'Unemi'
                    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
                    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                    title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    style1 = easyxf(num_format_str='D-MMM-YY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('exp_xls_post_part')
                    # ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    response = HttpResponse(content_type="application/ms-excel")
                    response[
                        'Content-Disposition'] = 'attachment; filename=Graduados' + random.randint(
                        1, 10000).__str__() + '.xls'
                    fech_ini = request.GET['fechainicio']
                    fech_fin = request.GET['fechafin']
                    cmbgenero = request.GET['cmbgenero']
                    carrera = request.GET['carrera']
                    columns = [
                        (u"tipoDocumentoId", 4000),
                        (u"numeroIdentificacion", 4000),
                        (u"primerApellido", 4000),
                        (u"segundoApellido", 4000),
                        (u"primerNombre", 4000),
                        (u"segundoNombre", 4000),
                        (u"sexoId", 1400),
                        (u"etniaId", 1400),
                        (u"fechaNacimiento", 5000),
                        (u"paisNacionalidadId", 2500),
                        (u"paisResidenciaId", 2500),
                        (u"provinciaResidenciaId", 2500),
                        (u"cantonResidenciaId", 2500),
                        (u"fechaInicioEstudios", 3000),
                        (u"fechaEgresamiento", 3000),
                        (u"duracion", 3000),
                        (u"tipoDuracionId", 3000),
                        (u"fechaActaGrado", 3000),
                        (u"numeroActaGrado", 3000),
                        (u"fechaRefrendacion", 3000),
                        (u"numeroRefrendacion", 3000),
                        (u"mecanismoTitulacionId", 6000),
                        (u"linkTesis", 3000),
                        (u"notaPromedioAcumulado", 3000),
                        (u"notaTrabajoTitulacion", 3000),
                        (u"reconocimientoEstudiosPrevios", 3000),
                        (u"institucionEstudiosPreviosId", 3000),
                        (u"carreraEstudiosPrevios", 3000),
                        (u"tiempoEstudiosReconocimiento", 3000),
                        (u"tipoDuracionReconocimiento", 3000),
                        (u"tituloBachiller", 15000),
                        (u"tipoColegioId", 3000),
                        (u"nombreRector", 10000),
                        (u"observaciones", 3000),
                        (u"carrera", 7000),
                    ]

                    row_num = 0
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    date_format = xlwt.XFStyle()
                    date_format.num_format_str = 'yyyy-mm-dd'
                    date_formatreverse = xlwt.XFStyle()
                    date_formatreverse.num_format_str = 'dd/mm/yyyy'
                    listadograduados = Graduado.objects.select_related().filter(inscripcion__persona__sexo__id=cmbgenero,
                                                                                fechaactagrado__gte=fech_ini,
                                                                                fechaactagrado__lte=fech_fin,
                                                                                status=True, inscripcion__carrera__id=carrera)
                    row_num = 1
                    i = 0
                    for listado in listadograduados:
                        campo1 = 0
                        actualizarpromediofinal(listado)
                        if listado.inscripcion.persona.cedula:
                            campo1 = 1
                        else:
                            campo1 = 2
                        campo2 = listado.inscripcion.persona.cedula
                        campo3 = remover_caracteres_tildes_unicode(listado.inscripcion.persona.apellido1)
                        campo4 = remover_caracteres_tildes_unicode(listado.inscripcion.persona.apellido2)
                        cadena = listado.inscripcion.persona.nombres.split(' ')
                        campo5 = remover_caracteres_tildes_unicode(cadena[0])
                        nombresegundo = ''
                        try:
                            if cadena[1]:
                                nombresegundo = remover_caracteres_tildes_unicode(cadena[1])
                        except:
                            pass
                        try:
                            if cadena[2]:
                                nombresegundo = remover_caracteres_tildes_unicode(cadena[1]) + ' ' + remover_caracteres_tildes_unicode(cadena[2])
                        except:
                            pass
                        try:
                            if cadena[3]:
                                nombresegundo = remover_caracteres_tildes_unicode(cadena[1]) + ' ' + remover_caracteres_tildes_unicode(cadena[2]) + ' ' + remover_caracteres_tildes_unicode(cadena[3])
                        except:
                            pass
                        try:
                            if cadena[4]:
                                nombresegundo = remover_caracteres_tildes_unicode(cadena[1]) + ' ' + remover_caracteres_tildes_unicode(cadena[2]) + ' ' + remover_caracteres_tildes_unicode(cadena[3]) + ' ' + remover_caracteres_tildes_unicode(cadena[4])
                        except:
                            pass
                        campo6 = nombresegundo
                        campo7 = 0
                        if listado.inscripcion.persona.sexo_id == 1:
                            campo7 = 2
                        if listado.inscripcion.persona.sexo_id == 2:
                            campo7 = 1
                        campo8 = listado.inscripcion.persona.mi_perfil().raza_id
                        campo9 = listado.inscripcion.persona.nacimiento
                        campo10 = 0
                        campo11 = 0
                        campo12 = 0
                        campo13 = 0
                        if listado.inscripcion.persona.paisnacimiento:
                            campo10 = listado.inscripcion.persona.paisnacimiento.codigosniese
                        if listado.inscripcion.persona.pais:
                            campo11 = listado.inscripcion.persona.pais.codigosniese
                        if listado.inscripcion.persona.provincia:
                            if listado.inscripcion.persona.provincia.codigosniese:
                                campo12 = listado.inscripcion.persona.provincia.codigosniese
                        if listado.inscripcion.persona.canton:
                            if listado.inscripcion.persona.canton.codigosniese:
                                campo13 = listado.inscripcion.persona.canton.codigosniese
                        campo14 = listado.inscripcion.fechainicioprimernivel
                        campo15 = None
                        if (listado.inscripcion.egresado_set.filter(status=True)).exists():
                            campo15 = listado.inscripcion.egresado_set.filter(status=True)[0].fechaegreso
                        campo16 = 8
                        campo17 = 2
                        campo18 = listado.fechaactagrado
                        campo19 = listado.numeroactagrado
                        campo20 = listado.fecharefrendacion
                        campo21 = listado.numeroactagrado
                        campo22 = ''
                        if listado.codigomecanismotitulacion:
                            campo22 = listado.codigomecanismotitulacion.codigosniese

                        # Link de tesis
                        campo23 = ''
                        if MatriculaTitulacion.objects.filter(Q(inscripcion=listado.inscripcion), (Q(estado=1) | Q(estado=10) | Q(estado=9))):
                            matriculatitulacion = MatriculaTitulacion.objects.filter(Q(inscripcion=listado.inscripcion), (Q(estado=1) | Q(estado=10) | Q(estado=9))).order_by('-id')[0]
                            if ComplexivoGrupoTematica.objects.filter(status=True, complexivodetallegrupo__status=True, complexivodetallegrupo__matricula=matriculatitulacion,activo=True).exists():
                                grupotitulacion = ComplexivoGrupoTematica.objects.filter(status=True, complexivodetallegrupo__status=True, complexivodetallegrupo__matricula=matriculatitulacion,activo=True).order_by('-id')[0]
                                if grupotitulacion.complexivopropuestapractica_set.filter(estado=2, status=True):
                                    if grupotitulacion.estadoarchivofinalgrupo == 2:
                                        campo23 = 'https://sga.unemi.edu.ec/media/' + grupotitulacion.archivofinalgrupo.__str__()
                                    else:
                                        grupopropuesta = grupotitulacion.complexivopropuestapractica_set.filter(estado=2, status=True).order_by('-id')[0]
                                        if grupopropuesta.complexivopropuestapracticaarchivo_set.filter(tipo=1, status=True).exists():
                                            campo23 = 'https://sga.unemi.edu.ec/media/' + grupopropuesta.complexivopropuestapracticaarchivo_set.get(tipo=1, status=True).archivo.__str__()
                        campo24 = str(listado.inscripcion.promedio_record())
                        campo25 = str(listado.promediotitulacion)
                        campo26 = 2
                        campo27 = ''
                        campo28 = ''
                        campo29 = ''
                        campo30 = ''
                        if listado.inscripcion.especialidad:
                            campo31 = listado.inscripcion.especialidad.nombre
                        else:
                            campo31 = ''
                        # if listado.inscripcion.colegio:
                        #     campo32 = ''
                        # else:
                        if listado.inscripcion.unidadeducativa:
                            campo32 = listado.inscripcion.unidadeducativa.tipocolegio.id
                        else:
                            campo32 = ''
                        # campo32 = ''
                        campo33 = 'JORGE FABRICIO GUEVARA VIEJO'
                        campo34 = ''
                        campo35 = listado.inscripcion.carrera.nombre + (" MENCION " + listado.inscripcion.carrera.mencion if listado.inscripcion.carrera.mencion else "")
                        i += 1
                        # ws.write(row_num, 0, i, font_style2)
                        ws.write(row_num, 0, campo1, font_style2)
                        ws.write(row_num, 1, campo2, font_style2)
                        ws.write(row_num, 2, campo3, font_style2)
                        ws.write(row_num, 3, campo4, font_style2)
                        ws.write(row_num, 4, campo5, font_style2)
                        ws.write(row_num, 5, campo6, font_style2)
                        ws.write(row_num, 6, campo7, font_style2)
                        ws.write(row_num, 7, campo8, font_style2)
                        ws.write(row_num, 8, campo9, date_format)
                        ws.write(row_num, 9, campo10, font_style2)
                        ws.write(row_num, 10, campo11, font_style2)
                        ws.write(row_num, 11, campo12, font_style2)
                        ws.write(row_num, 12, campo13, font_style2)
                        ws.write(row_num, 13, campo14, date_format)
                        ws.write(row_num, 14, campo15, date_format)
                        ws.write(row_num, 15, campo16, font_style2)
                        ws.write(row_num, 16, campo17, font_style2)
                        ws.write(row_num, 17, campo18, date_format)
                        ws.write(row_num, 18, campo19, font_style2)
                        ws.write(row_num, 19, campo20, date_format)
                        ws.write(row_num, 20, campo21, font_style2)
                        ws.write(row_num, 21, campo22, font_style2)
                        ws.write(row_num, 22, campo23, font_style2)
                        ws.write(row_num, 23, campo24, font_style2)
                        ws.write(row_num, 24, campo25, font_style2)
                        ws.write(row_num, 25, campo26, font_style2)
                        ws.write(row_num, 26, campo27, font_style2)
                        ws.write(row_num, 27, campo28, font_style2)
                        ws.write(row_num, 28, campo29, font_style2)
                        ws.write(row_num, 29, campo30, font_style2)
                        ws.write(row_num, 30, campo31, font_style2)
                        ws.write(row_num, 31, campo32, font_style2)
                        ws.write(row_num, 32, campo33, font_style2)
                        ws.write(row_num, 33, campo34, font_style2)
                        ws.write(row_num, 34, campo35, font_style2)
                        row_num += 1
                    wb.save(response)
                    return response
                except Exception as ex:
                    pass

            elif action == 'excelgraduadosdip':
                try:
                    __author__ = 'Unemi'
                    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
                    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                    title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    style1 = easyxf(num_format_str='D-MMM-YY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('exp_xls_post_part')
                    # ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    response = HttpResponse(content_type="application/ms-excel")
                    response[
                        'Content-Disposition'] = 'attachment; filename=Graduados' + random.randint(
                        1, 10000).__str__() + '.xls'
                    fech_ini = request.GET['fechainicio']
                    fech_fin = request.GET['fechafin']
                    carrera = request.GET['carrera']
                    duracion = int(request.GET['duracion'])
                    columns = [
                        # (u"periodo", 4000),
                        (u"tipoDocumentoId", 4000),
                        (u"numeroIdentificacion", 4000),
                        (u"primerApellido", 4000),
                        (u"segundoApellido", 4000),
                        (u"primerNombre", 4000),
                        (u"segundoNombre", 4000),
                        (u"sexoId", 1400),
                        (u"etniaId", 1400),
                        (u"fechaNacimiento", 5000),
                        (u"paisNacionalidadId", 2500),
                        (u"paisResidenciaId", 2500),
                        (u"provinciaResidenciaId", 2500),
                        (u"cantonResidenciaId", 2500),
                        (u"fechaInicioEstudios", 3000),
                        (u"fechaEgresamiento", 3000),
                        (u"duracion", 3000),
                        (u"tipoDuracionId", 3000),
                        (u"tituloBachiller", 15000),
                        (u"tipoColegioId", 3000),
                        (u"reconocimientoEstudiosPrevios", 3000),
                        (u"institucionEstudiosPreviosId", 3000),
                        (u"carreraEstudiosPrevios", 3000),
                        (u"tiempoEstudiosReconocimiento", 3000),
                        (u"tipoDuracionReconocimiento", 3000),
                        (u"fechaActaGrado", 3000),
                        (u"numeroActaGrado", 3000),
                        (u"fechaRefrendacion", 3000),
                        (u"numeroRefrendacion", 3000),
                        (u"mecanismoTitulacionId", 6000),
                        (u"linkTesis", 3000),
                        (u"notaPromedioAcumulado", 3000),
                        (u"notaTrabajoTitulacion", 3000),
                        (u"nombreRector", 10000),
                        (u"observaciones", 3000),
                        (u"carrera", 7000),
                        (u"cohorte ", 7000),
                    ]

                    row_num = 0
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    date_format = xlwt.XFStyle()
                    date_format.num_format_str = 'yyyy-mm-dd'
                    date_formatreverse = xlwt.XFStyle()
                    date_formatreverse.num_format_str = 'dd/mm/yyyy'
                    listadograduados = TribunalTemaTitulacionPosgradoMatricula.objects.filter(tematitulacionposgradomatricula__matricula__inscripcion__carrera_id=carrera, tematitulacionposgradomatricula__estadotribunal=2, fechadefensa__gte=fech_ini, fechadefensa__lte=fech_fin )
                    row_num = 1
                    i = 0
                    for listado in listadograduados:
                        campo1 = 0
                        # actualizarpromediofinal(listado)
                        if listado.tematitulacionposgradomatricula.matricula.inscripcion.persona.cedula:
                            campo1 = 1
                        else:
                            campo1 = 2
                        campo2 = listado.tematitulacionposgradomatricula.matricula.inscripcion.persona.cedula
                        campo3 = remover_caracteres_tildes_unicode(listado.tematitulacionposgradomatricula.matricula.inscripcion.persona.apellido1)
                        campo4 = remover_caracteres_tildes_unicode(listado.tematitulacionposgradomatricula.matricula.inscripcion.persona.apellido2)
                        cadena = listado.tematitulacionposgradomatricula.matricula.inscripcion.persona.nombres.split(' ')
                        campo5 = remover_caracteres_tildes_unicode(cadena[0])
                        nombresegundo = ''
                        try:
                            if cadena[1]:
                                nombresegundo = remover_caracteres_tildes_unicode(cadena[1])
                        except:
                            pass
                        try:
                            if cadena[2]:
                                nombresegundo = remover_caracteres_tildes_unicode(cadena[1]) + ' ' + remover_caracteres_tildes_unicode(cadena[2])
                        except:
                            pass
                        try:
                            if cadena[3]:
                                nombresegundo = remover_caracteres_tildes_unicode(cadena[1]) + ' ' + remover_caracteres_tildes_unicode(cadena[2]) + ' ' + remover_caracteres_tildes_unicode(cadena[3])
                        except:
                            pass
                        try:
                            if cadena[4]:
                                nombresegundo = remover_caracteres_tildes_unicode(cadena[1]) + ' ' + remover_caracteres_tildes_unicode(cadena[2]) + ' ' + remover_caracteres_tildes_unicode(cadena[3]) + ' ' + remover_caracteres_tildes_unicode(cadena[4])
                        except:
                            pass
                        campo6 = nombresegundo
                        campo7 = 0
                        if listado.tematitulacionposgradomatricula.matricula.inscripcion.persona.sexo_id == 1:
                            campo7 = 2
                        if listado.tematitulacionposgradomatricula.matricula.inscripcion.persona.sexo_id == 2:
                            campo7 = 1
                        campo8 = listado.tematitulacionposgradomatricula.matricula.inscripcion.persona.mi_perfil().raza_id
                        campo9 = listado.tematitulacionposgradomatricula.matricula.inscripcion.persona.nacimiento
                        campo10 = 0
                        campo11 = 0
                        campo12 = 0
                        campo13 = 0
                        if listado.tematitulacionposgradomatricula.matricula.inscripcion.persona.paisnacimiento:
                            campo10 = listado.tematitulacionposgradomatricula.matricula.inscripcion.persona.paisnacimiento.codigosniese
                        if listado.tematitulacionposgradomatricula.matricula.inscripcion.persona.pais:
                            campo11 = listado.tematitulacionposgradomatricula.matricula.inscripcion.persona.pais.codigosniese
                        if listado.tematitulacionposgradomatricula.matricula.inscripcion.persona.provincia:
                            if listado.tematitulacionposgradomatricula.matricula.inscripcion.persona.provincia.codigosniese:
                                campo12 = listado.tematitulacionposgradomatricula.matricula.inscripcion.persona.provincia.codigosniese
                        if listado.tematitulacionposgradomatricula.matricula.inscripcion.persona.canton:
                            if listado.tematitulacionposgradomatricula.matricula.inscripcion.persona.canton.codigosniese:
                                campo13 = listado.tematitulacionposgradomatricula.matricula.inscripcion.persona.canton.codigosniese
                        campo14 = listado.tematitulacionposgradomatricula.matricula.inscripcion.fechainicioprimernivel
                        campo15 = listado.tematitulacionposgradomatricula.fechaacta
                        if (listado.tematitulacionposgradomatricula.matricula.inscripcion.egresado_set.filter(status=True)).exists():
                            campo15 = listado.tematitulacionposgradomatricula.matricula.inscripcion.egresado_set.filter(status=True)[0].fechaegreso
                        campo16 = duracion
                        campo17 = 2
                        campo18 = listado.fechadefensa
                        campo19 = listado.tematitulacionposgradomatricula.matricula.inscripcion.graduado_set.filter(status=True)[0].numeroactagrado
                        campo20 = listado.fechadefensa
                        campo21 = listado.tematitulacionposgradomatricula.matricula.inscripcion.graduado_set.filter(status=True)[0].numeroactagrado
                        campo22 = ''
                        if listado.tematitulacionposgradomatricula.mecanismotitulacionposgrado:
                            campo22 = listado.tematitulacionposgradomatricula.mecanismotitulacionposgrado.id

                        # Link de tesis
                        campo23 = ''
                        if MatriculaTitulacion.objects.filter(Q(inscripcion=listado.tematitulacionposgradomatricula.matricula.inscripcion), (Q(estado=1) | Q(estado=10) | Q(estado=9))):
                            matriculatitulacion = MatriculaTitulacion.objects.filter(Q(inscripcion=listado.tematitulacionposgradomatricula.matricula.inscripcion), (Q(estado=1) | Q(estado=10) | Q(estado=9))).order_by('-id')[0]
                            if ComplexivoGrupoTematica.objects.filter(status=True, complexivodetallegrupo__status=True, complexivodetallegrupo__matricula=matriculatitulacion).exists():
                                grupotitulacion = ComplexivoGrupoTematica.objects.filter(status=True, complexivodetallegrupo__status=True, complexivodetallegrupo__matricula=matriculatitulacion).order_by('-id')[0]
                                if grupotitulacion.complexivopropuestapractica_set.filter(estado=2, status=True):
                                    if grupotitulacion.estadoarchivofinalgrupo == 2:
                                        campo23 = 'https://sga.unemi.edu.ec/media/' + grupotitulacion.archivofinalgrupo.__str__()
                                    else:
                                        grupopropuesta = grupotitulacion.complexivopropuestapractica_set.filter(estado=2, status=True).order_by('-id')[0]
                                        if grupopropuesta.complexivopropuestapracticaarchivo_set.filter(tipo=1, status=True).exists():
                                            campo23 = 'https://sga.unemi.edu.ec/media/' + grupopropuesta.complexivopropuestapracticaarchivo_set.get(tipo=1, status=True).archivo.__str__()
                        campo24 = str(listado.tematitulacionposgradomatricula.matricula.inscripcion.graduado_set.filter(status=True)[0].notafinal)
                        campo25 = str(listado.tematitulacionposgradomatricula.matricula.inscripcion.graduado_set.filter(status=True)[0].promediotitulacion)
                        campo26 = 2
                        campo27 = ''
                        campo28 = ''
                        campo29 = ''
                        campo30 = ''
                        if listado.tematitulacionposgradomatricula.matricula.inscripcion.especialidad:
                            campo31 = listado.tematitulacionposgradomatricula.matricula.inscripcion.especialidad.nombre
                        else:
                            campo31 = ''
                        # if listado.inscripcion.colegio:
                        #     campo32 = ''
                        # else:
                        if listado.tematitulacionposgradomatricula.matricula.inscripcion.unidadeducativa:
                            campo32 = listado.tematitulacionposgradomatricula.matricula.inscripcion.unidadeducativa.tipocolegio.id
                        else:
                            campo32 = ''
                        # campo32 = ''
                        campo33 = 'JORGE FABRICIO GUEVARA VIEJO'
                        campo34 = ''
                        campo35 = listado.tematitulacionposgradomatricula.matricula.inscripcion.carrera.nombre + (" MENCION " + listado.tematitulacionposgradomatricula.matricula.inscripcion.carrera.mencion if listado.tematitulacionposgradomatricula.matricula.inscripcion.carrera.mencion else "")
                        campo36 = listado.tematitulacionposgradomatricula.convocatoria.periodo.nombre
                        i += 1
                        # ws.write(row_num, 0, i, font_style2)
                        ws.write(row_num, 0, campo1, font_style2)
                        ws.write(row_num, 1, campo2, font_style2)
                        ws.write(row_num, 2, campo3, font_style2)
                        ws.write(row_num, 3, campo4, font_style2)
                        ws.write(row_num, 4, campo5, font_style2)
                        ws.write(row_num, 5, campo6, font_style2)
                        ws.write(row_num, 6, campo7, font_style2)
                        ws.write(row_num, 7, campo8, font_style2)
                        ws.write(row_num, 8, campo9, date_format)
                        ws.write(row_num, 9, campo10, font_style2)
                        ws.write(row_num, 10, campo11, font_style2)
                        ws.write(row_num, 11, campo12, font_style2)
                        ws.write(row_num, 12, campo13, font_style2)
                        ws.write(row_num, 13, campo14, date_format)
                        ws.write(row_num, 14, campo15, date_format)
                        ws.write(row_num, 15, campo16, font_style2)
                        ws.write(row_num, 16, campo17, font_style2)
                        ws.write(row_num, 17, campo31, font_style2)
                        ws.write(row_num, 18, campo32, font_style2)
                        ws.write(row_num, 19, campo26, font_style2)
                        ws.write(row_num, 20, campo27, font_style2)
                        ws.write(row_num, 21, campo28, font_style2)
                        ws.write(row_num, 22, campo29, font_style2)
                        ws.write(row_num, 23, campo30, font_style2)
                        ws.write(row_num, 24, campo18, date_format)
                        ws.write(row_num, 25, campo19, font_style2)
                        ws.write(row_num, 26, campo20, date_format)
                        ws.write(row_num, 27, campo21, font_style2)
                        ws.write(row_num, 28, campo22, font_style2)
                        ws.write(row_num, 29, campo23, font_style2)
                        ws.write(row_num, 30, campo24, font_style2)
                        ws.write(row_num, 31, campo25, font_style2)
                        ws.write(row_num, 32, campo33, font_style2)
                        ws.write(row_num, 33, campo34, font_style2)
                        ws.write(row_num, 34, campo35, font_style2)
                        ws.write(row_num, 35, campo36, font_style2)
                        row_num += 1
                    listadograduados_pareja = TribunalTemaTitulacionPosgradoMatricula.objects.filter(tematitulacionposgradomatriculacabecera__convocatoria__carrera_id=carrera, fechadefensa__gte=fech_ini, fechadefensa__lte=fech_fin )
                    #en pareja
                    for listado in listadograduados_pareja:
                        for participante in listado.tematitulacionposgradomatriculacabecera.obtener_parejas():
                            if participante.estadotribunal== 2:
                                campo1 = 0
                                # actualizarpromediofinal(listado)
                                if participante.matricula.inscripcion.persona.cedula:
                                    campo1 = 1
                                else:
                                    campo1 = 2
                                campo2 = participante.matricula.inscripcion.persona.cedula
                                campo3 = remover_caracteres_tildes_unicode(
                                    participante.matricula.inscripcion.persona.apellido1)
                                campo4 = remover_caracteres_tildes_unicode(
                                    participante.matricula.inscripcion.persona.apellido2)
                                cadena = participante.matricula.inscripcion.persona.nombres.split(
                                    ' ')
                                campo5 = remover_caracteres_tildes_unicode(cadena[0])
                                nombresegundo = ''
                                try:
                                    if cadena[1]:
                                        nombresegundo = remover_caracteres_tildes_unicode(cadena[1])
                                except:
                                    pass
                                try:
                                    if cadena[2]:
                                        nombresegundo = remover_caracteres_tildes_unicode(
                                            cadena[1]) + ' ' + remover_caracteres_tildes_unicode(cadena[2])
                                except:
                                    pass
                                try:
                                    if cadena[3]:
                                        nombresegundo = remover_caracteres_tildes_unicode(
                                            cadena[1]) + ' ' + remover_caracteres_tildes_unicode(
                                            cadena[2]) + ' ' + remover_caracteres_tildes_unicode(cadena[3])
                                except:
                                    pass
                                try:
                                    if cadena[4]:
                                        nombresegundo = remover_caracteres_tildes_unicode(
                                            cadena[1]) + ' ' + remover_caracteres_tildes_unicode(
                                            cadena[2]) + ' ' + remover_caracteres_tildes_unicode(
                                            cadena[3]) + ' ' + remover_caracteres_tildes_unicode(cadena[4])
                                except:
                                    pass
                                campo6 = nombresegundo
                                campo7 = 0
                                if participante.matricula.inscripcion.persona.sexo_id == 1:
                                    campo7 = 2
                                if participante.matricula.inscripcion.persona.sexo_id == 2:
                                    campo7 = 1
                                campo8 = participante.matricula.inscripcion.persona.mi_perfil().raza_id
                                campo9 = participante.matricula.inscripcion.persona.nacimiento
                                campo10 = 0
                                campo11 = 0
                                campo12 = 0
                                campo13 = 0
                                if participante.matricula.inscripcion.persona.paisnacimiento:
                                    campo10 = participante.matricula.inscripcion.persona.paisnacimiento.codigosniese
                                if participante.matricula.inscripcion.persona.pais:
                                    campo11 = participante.matricula.inscripcion.persona.pais.codigosniese
                                if participante.matricula.inscripcion.persona.provincia:
                                    if participante.matricula.inscripcion.persona.provincia.codigosniese:
                                        campo12 = participante.matricula.inscripcion.persona.provincia.codigosniese
                                if participante.matricula.inscripcion.persona.canton:
                                    if participante.matricula.inscripcion.persona.canton.codigosniese:
                                        campo13 = participante.matricula.inscripcion.persona.canton.codigosniese
                                campo14 = participante.matricula.inscripcion.fechainicioprimernivel
                                campo15 = participante.fechaacta
                                if (participante.matricula.inscripcion.egresado_set.filter(
                                        status=True)).exists():
                                    campo15 = \
                                        participante.matricula.inscripcion.egresado_set.filter(
                                            status=True)[0].fechaegreso
                                campo16 = duracion
                                campo17 = 2
                                campo18 = listado.fechadefensa
                                campo19 = participante.matricula.inscripcion.graduado_set.filter(
                                    status=True)[0].numeroactagrado
                                campo20 = listado.fechadefensa
                                campo21 = participante.matricula.inscripcion.graduado_set.filter(
                                    status=True)[0].numeroactagrado
                                campo22 = ''
                                if participante.mecanismotitulacionposgrado:
                                    campo22 = participante.mecanismotitulacionposgrado.id

                                # Link de tesis
                                campo23 = ''
                                if MatriculaTitulacion.objects.filter(
                                        Q(inscripcion=participante.matricula.inscripcion),
                                        (Q(estado=1) | Q(estado=10) | Q(estado=9))):
                                    matriculatitulacion = MatriculaTitulacion.objects.filter(
                                        Q(inscripcion=participante.matricula.inscripcion),
                                        (Q(estado=1) | Q(estado=10) | Q(estado=9))).order_by('-id')[0]
                                    if ComplexivoGrupoTematica.objects.filter(status=True,
                                                                              complexivodetallegrupo__status=True,
                                                                              complexivodetallegrupo__matricula=matriculatitulacion).exists():
                                        grupotitulacion = ComplexivoGrupoTematica.objects.filter(status=True,
                                                                                                 complexivodetallegrupo__status=True,
                                                                                                 complexivodetallegrupo__matricula=matriculatitulacion).order_by(
                                            '-id')[0]
                                        if grupotitulacion.complexivopropuestapractica_set.filter(estado=2, status=True):
                                            if grupotitulacion.estadoarchivofinalgrupo == 2:
                                                campo23 = 'https://sga.unemi.edu.ec/media/' + grupotitulacion.archivofinalgrupo.__str__()
                                            else:
                                                grupopropuesta = \
                                                    grupotitulacion.complexivopropuestapractica_set.filter(estado=2,
                                                                                                           status=True).order_by(
                                                        '-id')[0]
                                                if grupopropuesta.complexivopropuestapracticaarchivo_set.filter(tipo=1,
                                                                                                                status=True).exists():
                                                    campo23 = 'https://sga.unemi.edu.ec/media/' + grupopropuesta.complexivopropuestapracticaarchivo_set.get(
                                                        tipo=1, status=True).archivo.__str__()
                                campo24 = str(
                                    participante.matricula.inscripcion.graduado_set.filter(
                                        status=True)[0].notafinal)
                                campo25 = str(
                                    participante.matricula.inscripcion.graduado_set.filter(
                                        status=True)[0].promediotitulacion)
                                campo26 = 2
                                campo27 = ''
                                campo28 = ''
                                campo29 = ''
                                campo30 = ''
                                if participante.matricula.inscripcion.especialidad:
                                    campo31 = participante.matricula.inscripcion.especialidad.nombre
                                else:
                                    campo31 = ''
                                # if listado.inscripcion.colegio:
                                #     campo32 = ''
                                # else:
                                if participante.matricula.inscripcion.unidadeducativa:
                                    campo32 = participante.matricula.inscripcion.unidadeducativa.tipocolegio.id
                                else:
                                    campo32 = ''
                                # campo32 = ''
                                campo33 = 'JORGE FABRICIO GUEVARA VIEJO'
                                campo34 = ''
                                campo35 = participante.matricula.inscripcion.carrera.nombre + (
                                    " MENCION " + participante.matricula.inscripcion.carrera.mencion if participante.matricula.inscripcion.carrera.mencion else "")
                                campo36 = participante.convocatoria.periodo.nombre
                                i += 1
                                # ws.write(row_num, 0, i, font_style2)
                                ws.write(row_num, 0, campo1, font_style2)
                                ws.write(row_num, 1, campo2, font_style2)
                                ws.write(row_num, 2, campo3, font_style2)
                                ws.write(row_num, 3, campo4, font_style2)
                                ws.write(row_num, 4, campo5, font_style2)
                                ws.write(row_num, 5, campo6, font_style2)
                                ws.write(row_num, 6, campo7, font_style2)
                                ws.write(row_num, 7, campo8, font_style2)
                                ws.write(row_num, 8, campo9, date_format)
                                ws.write(row_num, 9, campo10, font_style2)
                                ws.write(row_num, 10, campo11, font_style2)
                                ws.write(row_num, 11, campo12, font_style2)
                                ws.write(row_num, 12, campo13, font_style2)
                                ws.write(row_num, 13, campo14, date_format)
                                ws.write(row_num, 14, campo15, date_format)
                                ws.write(row_num, 15, campo16, font_style2)
                                ws.write(row_num, 16, campo17, font_style2)
                                ws.write(row_num, 17, campo31, font_style2)
                                ws.write(row_num, 18, campo32, font_style2)
                                ws.write(row_num, 19, campo26, font_style2)
                                ws.write(row_num, 20, campo27, font_style2)
                                ws.write(row_num, 21, campo28, font_style2)
                                ws.write(row_num, 22, campo29, font_style2)
                                ws.write(row_num, 23, campo30, font_style2)
                                ws.write(row_num, 24, campo18, date_format)
                                ws.write(row_num, 25, campo19, font_style2)
                                ws.write(row_num, 26, campo20, date_format)
                                ws.write(row_num, 27, campo21, font_style2)
                                ws.write(row_num, 28, campo22, font_style2)
                                ws.write(row_num, 29, campo23, font_style2)
                                ws.write(row_num, 30, campo24, font_style2)
                                ws.write(row_num, 31, campo25, font_style2)
                                ws.write(row_num, 32, campo33, font_style2)
                                ws.write(row_num, 33, campo34, font_style2)
                                ws.write(row_num, 34, campo35, font_style2)
                                ws.write(row_num, 35, campo36, font_style2)
                                row_num += 1
                    #complexivo

                    listadograduados_complexivo = TemaTitulacionPosgradoMatricula.objects.filter(matricula__inscripcion__carrera_id=carrera, actacerrada= True, aprobado=True, estado_acta_firma = 4)

                    for listado in listadograduados_complexivo:
                        if listado.matricula.inscripcion.es_graduado():
                            campo1 = 0
                            # actualizarpromediofinal(listado)
                            if listado.matricula.inscripcion.persona.cedula:
                                campo1 = 1
                            else:
                                campo1 = 2
                            campo2 = listado.matricula.inscripcion.persona.cedula
                            campo3 = remover_caracteres_tildes_unicode(
                                listado.matricula.inscripcion.persona.apellido1)
                            campo4 = remover_caracteres_tildes_unicode(
                                listado.matricula.inscripcion.persona.apellido2)
                            cadena = listado.matricula.inscripcion.persona.nombres.split(
                                ' ')
                            campo5 = remover_caracteres_tildes_unicode(cadena[0])
                            nombresegundo = ''
                            try:
                                if cadena[1]:
                                    nombresegundo = remover_caracteres_tildes_unicode(cadena[1])
                            except:
                                pass
                            try:
                                if cadena[2]:
                                    nombresegundo = remover_caracteres_tildes_unicode(
                                        cadena[1]) + ' ' + remover_caracteres_tildes_unicode(cadena[2])
                            except:
                                pass
                            try:
                                if cadena[3]:
                                    nombresegundo = remover_caracteres_tildes_unicode(
                                        cadena[1]) + ' ' + remover_caracteres_tildes_unicode(
                                        cadena[2]) + ' ' + remover_caracteres_tildes_unicode(cadena[3])
                            except:
                                pass
                            try:
                                if cadena[4]:
                                    nombresegundo = remover_caracteres_tildes_unicode(
                                        cadena[1]) + ' ' + remover_caracteres_tildes_unicode(
                                        cadena[2]) + ' ' + remover_caracteres_tildes_unicode(
                                        cadena[3]) + ' ' + remover_caracteres_tildes_unicode(cadena[4])
                            except:
                                pass
                            campo6 = nombresegundo
                            campo7 = 0
                            if listado.matricula.inscripcion.persona.sexo_id == 1:
                                campo7 = 2
                            if listado.matricula.inscripcion.persona.sexo_id == 2:
                                campo7 = 1
                            campo8 = listado.matricula.inscripcion.persona.mi_perfil().raza_id
                            campo9 = listado.matricula.inscripcion.persona.nacimiento
                            campo10 = 0
                            campo11 = 0
                            campo12 = 0
                            campo13 = 0
                            if listado.matricula.inscripcion.persona.paisnacimiento:
                                campo10 = listado.matricula.inscripcion.persona.paisnacimiento.codigosniese
                            if listado.matricula.inscripcion.persona.pais:
                                campo11 = listado.matricula.inscripcion.persona.pais.codigosniese
                            if listado.matricula.inscripcion.persona.provincia:
                                if listado.matricula.inscripcion.persona.provincia.codigosniese:
                                    campo12 = listado.matricula.inscripcion.persona.provincia.codigosniese
                            if listado.matricula.inscripcion.persona.canton:
                                if listado.matricula.inscripcion.persona.canton.codigosniese:
                                    campo13 = listado.matricula.inscripcion.persona.canton.codigosniese
                            campo14 = listado.matricula.inscripcion.fechainicioprimernivel
                            campo15 = listado.fechaacta
                            if (listado.matricula.inscripcion.egresado_set.filter(
                                    status=True)).exists():
                                campo15 = listado.matricula.inscripcion.egresado_set.filter(
                                    status=True)[0].fechaegreso
                            campo16 = duracion
                            campo17 = 2
                            campo18 = listado.obtener_fecha_examen_complexivo()
                            campo19 = listado.matricula.inscripcion.graduado_set.filter(status=True)[0].numeroactagrado
                            campo20 = listado.obtener_fecha_examen_complexivo()
                            campo21 = listado.matricula.inscripcion.graduado_set.filter(status=True)[0].numeroactagrado
                            campo22 = ''
                            if listado.mecanismotitulacionposgrado:
                                campo22 = listado.mecanismotitulacionposgrado.id

                            # Link de tesis
                            campo23 = ''
                            if MatriculaTitulacion.objects.filter(
                                    Q(inscripcion=listado.matricula.inscripcion),
                                    (Q(estado=1) | Q(estado=10) | Q(estado=9))):
                                matriculatitulacion = MatriculaTitulacion.objects.filter(
                                    Q(inscripcion=listado.matricula.inscripcion),
                                    (Q(estado=1) | Q(estado=10) | Q(estado=9))).order_by('-id')[0]
                                if ComplexivoGrupoTematica.objects.filter(status=True, complexivodetallegrupo__status=True,
                                                                          complexivodetallegrupo__matricula=matriculatitulacion).exists():
                                    grupotitulacion = \
                                        ComplexivoGrupoTematica.objects.filter(status=True, complexivodetallegrupo__status=True,
                                                                               complexivodetallegrupo__matricula=matriculatitulacion).order_by(
                                            '-id')[0]
                                    if grupotitulacion.complexivopropuestapractica_set.filter(estado=2, status=True):
                                        if grupotitulacion.estadoarchivofinalgrupo == 2:
                                            campo23 = 'https://sga.unemi.edu.ec/media/' + grupotitulacion.archivofinalgrupo.__str__()
                                        else:
                                            grupopropuesta = \
                                                grupotitulacion.complexivopropuestapractica_set.filter(estado=2,
                                                                                                       status=True).order_by(
                                                    '-id')[0]
                                            if grupopropuesta.complexivopropuestapracticaarchivo_set.filter(tipo=1,
                                                                                                            status=True).exists():
                                                campo23 = 'https://sga.unemi.edu.ec/media/' + grupopropuesta.complexivopropuestapracticaarchivo_set.get(
                                                    tipo=1, status=True).archivo.__str__()
                            campo24 = str(listado.matricula.inscripcion.graduado_set.filter(
                                status=True)[0].notafinal)
                            campo25 = str(listado.matricula.inscripcion.graduado_set.filter(
                                status=True)[0].promediotitulacion)
                            campo26 = 2
                            campo27 = ''
                            campo28 = ''
                            campo29 = ''
                            campo30 = ''
                            if listado.matricula.inscripcion.especialidad:
                                campo31 = listado.matricula.inscripcion.especialidad.nombre
                            else:
                                campo31 = ''
                            # if listado.inscripcion.colegio:
                            #     campo32 = ''
                            # else:
                            if listado.matricula.inscripcion.unidadeducativa:
                                campo32 = listado.matricula.inscripcion.unidadeducativa.tipocolegio.id
                            else:
                                campo32 = ''
                            # campo32 = ''
                            campo33 = 'JORGE FABRICIO GUEVARA VIEJO'
                            campo34 = ''
                            campo35 = listado.matricula.inscripcion.carrera.nombre + (
                                " MENCION " + listado.matricula.inscripcion.carrera.mencion if listado.matricula.inscripcion.carrera.mencion else "")
                            campo36 = listado.convocatoria.periodo.nombre
                            i += 1
                            # ws.write(row_num, 0, i, font_style2)
                            ws.write(row_num, 0, campo1, font_style2)
                            ws.write(row_num, 1, campo2, font_style2)
                            ws.write(row_num, 2, campo3, font_style2)
                            ws.write(row_num, 3, campo4, font_style2)
                            ws.write(row_num, 4, campo5, font_style2)
                            ws.write(row_num, 5, campo6, font_style2)
                            ws.write(row_num, 6, campo7, font_style2)
                            ws.write(row_num, 7, campo8, font_style2)
                            ws.write(row_num, 8, campo9, date_format)
                            ws.write(row_num, 9, campo10, font_style2)
                            ws.write(row_num, 10, campo11, font_style2)
                            ws.write(row_num, 11, campo12, font_style2)
                            ws.write(row_num, 12, campo13, font_style2)
                            ws.write(row_num, 13, campo14, date_format)
                            ws.write(row_num, 14, campo15, date_format)
                            ws.write(row_num, 15, campo16, font_style2)
                            ws.write(row_num, 16, campo17, font_style2)
                            ws.write(row_num, 17, campo31, font_style2)
                            ws.write(row_num, 18, campo32, font_style2)
                            ws.write(row_num, 19, campo26, font_style2)
                            ws.write(row_num, 20, campo27, font_style2)
                            ws.write(row_num, 21, campo28, font_style2)
                            ws.write(row_num, 22, campo29, font_style2)
                            ws.write(row_num, 23, campo30, font_style2)
                            ws.write(row_num, 24, campo18, date_format)
                            ws.write(row_num, 25, campo19, font_style2)
                            ws.write(row_num, 26, campo20, date_format)
                            ws.write(row_num, 27, campo21, font_style2)
                            ws.write(row_num, 28, campo22, font_style2)
                            ws.write(row_num, 29, campo23, font_style2)
                            ws.write(row_num, 30, campo24, font_style2)
                            ws.write(row_num, 31, campo25, font_style2)
                            ws.write(row_num, 32, campo33, font_style2)
                            ws.write(row_num, 33, campo34, font_style2)
                            ws.write(row_num, 34, campo35, font_style2)
                            ws.write(row_num, 35, campo36, font_style2)
                            row_num += 1

                    wb.save(response)
                    return response
                except Exception as ex:
                    pass

            elif action == 'excelgraduadosrural':
                try:
                    __author__ = 'Unemi'
                    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
                    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                    title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    style1 = easyxf(num_format_str='D-MMM-YY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('exp_xls_post_part')
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=Graduados' + random.randint(1, 10000).__str__() + '.xls'
                    fech_ini = request.GET['fechainicio']
                    fech_fin = request.GET['fechafin']
                    carrera = request.GET['carrera']
                    columns = [
                        (u"N.", 4000),
                        (u"ZONA", 4000),
                        (u"UNIVERSIDAD", 4000),
                        (u"CARRERA", 4000),
                        (u"TOTAL GRADUADOS", 4000),
                        (u"APELLIDOS Y NOMBRES", 15000),
                        (u"NUMERO DE DOCUMENTO DE IDENTIFICACION", 4000),
                        (u"NOTA DE GRADO", 2000),
                        (u"PROMOCION", 5000),
                        (u"OBSERVACION DE LA CTZR", 2500),
                    ]
                    row_num = 0
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    date_format = xlwt.XFStyle()
                    date_format.num_format_str = 'yyyy-mm-dd'
                    date_formatreverse = xlwt.XFStyle()
                    date_formatreverse.num_format_str = 'dd/mm/yyyy'
                    listadograduados = Graduado.objects.select_related().filter(fechaactagrado__gte=fech_ini, fechaactagrado__lte=fech_fin, status=True, inscripcion__carrera__id=carrera).order_by('-notagraduacion')
                    totalgraduados = listadograduados.count()
                    row_num = 1
                    i = 0
                    for listado in listadograduados:
                        campo2 = 'ZONA 5'
                        textoidentidad = ''
                        if listado.inscripcion.persona.cedula:
                            textoidentidad = listado.inscripcion.persona.cedula
                        else:
                            if listado.inscripcion.persona.pasaporte:
                                textoidentidad = listado.inscripcion.persona.pasaporte
                        campo7 = textoidentidad
                        campo8 = listado.notagraduacion
                        campo3 = 'UNIVERSIDAD ESTATAL DE MLAGRO'
                        primerapellido = remover_caracteres_tildes_unicode(listado.inscripcion.persona.apellido1)
                        segundoapellido = remover_caracteres_tildes_unicode(listado.inscripcion.persona.apellido2)
                        nombres = listado.inscripcion.persona.nombres
                        campo6 = primerapellido + ' ' + segundoapellido + ' ' + nombres
                        campo4 = listado.inscripcion.carrera.nombre + (" MENCION " + listado.inscripcion.carrera.mencion if listado.inscripcion.carrera.mencion else "")
                        i += 1
                        campo9 = ''
                        ws.write(row_num, 0, i, font_style2)
                        ws.write(row_num, 1, campo2, font_style2)
                        ws.write(row_num, 2, campo3, font_style2)
                        ws.write(row_num, 3, campo4, font_style2)
                        ws.write(row_num, 4, totalgraduados, font_style2)
                        ws.write(row_num, 5, campo6, font_style2)
                        ws.write(row_num, 6, campo7, font_style2)
                        ws.write(row_num, 7, campo8, font_style2)
                        ws.write(row_num, 8, campo9, date_format)
                        row_num += 1
                    wb.save(response)
                    return response
                except Exception as ex:
                    pass

            elif action == 'listadoactasgraduados':
                try:
                    data['title'] = u'Listado de actas'
                    data['graduado'] = graduado = Graduado.objects.get(status=True, pk=int(request.GET['id']))
                    # data['listadoactas'] = TipoActa.objects.filter(status=True, graduado_id=int(request.GET['id']))
                    return render(request, "graduados/actasgraduados.html", data)
                except Exception as ex:
                    pass

            elif action == 'listadoactasgrado':
                try:
                    url_vars, search, = f'&action={action}', request.GET.get('s', '').strip()
                    data['title'] = u'Listado de actas grado'
                    data['horaactual'] = horaactual
                    listado = persona.tipoactafirma_set.filter(tipoacta__tipo=5, firmado=False, turnofirmar=True, status=True).order_by('-tipoacta__graduado__fechagraduado','tipoacta__graduado__inscripcion__persona__apellido1')
                    data['carreras'] = Carrera.objects.filter(pk__in=listado.values_list('tipoacta__graduado__inscripcion__carrera_id', flat=True).distinct(), status=True).order_by('nombre')

                    if 's' in request.GET:
                        search = request.GET['s']
                        url_vars += f"&s={search}"
                        ss = search.split(' ')
                        if len(ss) == 1:
                            listado = listado.filter(Q(tipoacta__graduado__inscripcion__persona__nombres__icontains=search) |
                                                     Q(tipoacta__graduado__inscripcion__persona__apellido1__icontains=search) |
                                                     Q(tipoacta__graduado__inscripcion__persona__apellido2__icontains=search) |
                                                     Q(tipoacta__graduado__inscripcion__persona__cedula__icontains=search) |
                                                     Q(tipoacta__graduado__inscripcion__persona__pasaporte__icontains=search))
                        else:
                            listado = listado.filter(Q(tipoacta__graduado__inscripcion__persona__apellido1__icontains=ss[0]) &
                                                     Q(tipoacta__graduado__inscripcion__persona__apellido2__icontains=ss[1]))
                    idc = 0
                    if 'idc' in request.GET:
                        idc = int(request.GET['idc'])
                        url_vars += f"&idc={idc}"
                        if idc > 0:
                            listado = listado.filter(tipoacta__graduado__inscripcion__carrera__id=idc)
                    fechini = None
                    fechfin = None
                    if 'fechini' in request.GET and 'fechfin' in request.GET:
                        fechini = request.GET['fechini']
                        fechfin = request.GET['fechfin']
                        if fechini and fechfin:
                            url_vars += f"&fechini={fechini}"
                            url_vars += f"&fechfin={fechfin}"
                            listado = listado.filter(tipoacta__graduado__fechagraduado__gte=fechini, tipoacta__graduado__fechagraduado__lte=fechfin)
                    numerofilas = 25
                    paging = MiPaginador(listado, numerofilas)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                            if p == 1:
                                numerofilasguiente = numerofilas
                            else:
                                numerofilasguiente = numerofilas * (p - 1)
                        else:
                            p = paginasesion
                            if p == 1:
                                numerofilasguiente = numerofilas
                            else:
                                numerofilasguiente = numerofilas * (p - 1)
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['numeropagina'] = p
                    data['numerofilasguiente'] = numerofilasguiente
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['idc'] = idc
                    data['fechini'] = fechini
                    data['fechfin'] = fechfin
                    data['search'] = search if search else ""
                    data['url_vars'] = url_vars
                    data['listado'] = page.object_list
                    request.session['viewactivoth'] = ['pendienteactagrado', 'listadoactasgrado']
                    return render(request, "graduados/listadoalumnosfirmar.html", data)
                except Exception as ex:
                    pass

            elif action == 'listadofirmadasactasgrado':
                try:
                    puede_eliminar_actas_firmadas = variable_valor('PUEDE_ELIMINAR_ACTAS_FIRMADAS')
                    url_vars, search, = f'&action={action}', request.GET.get('s', '').strip()
                    data['title'] = u'Actas grado firmadas'
                    data['horaactual'] = horaactual
                    puedeeliminar = True
                    if user.is_superuser:
                        puedeeliminar = False
                        listado = TipoActaFirma.objects.filter(tipoacta__tipo=5, firmado=True, status=True, orden=1).order_by('-tipoacta__graduado__fechagraduado','tipoacta__graduado__inscripcion__persona__apellido1')
                    else:
                        listado = persona.tipoactafirma_set.filter(tipoacta__tipo=5, firmado=True, status=True).order_by('-tipoacta__graduado__fechagraduado','tipoacta__graduado__inscripcion__persona__apellido1')
                    data['puedeeliminar'] = puedeeliminar
                    data['puede_eliminar_actas_firmadas'] = puede_eliminar_actas_firmadas
                    data['carreras'] = Carrera.objects.filter(pk__in=listado.values_list('tipoacta__graduado__inscripcion__carrera_id', flat=True).distinct(), status=True).order_by('nombre')
                    if 's' in request.GET:
                        search = request.GET['s']
                        url_vars += f"&s={search}"
                        ss = search.split(' ')
                        if len(ss) == 1:
                            listado = listado.filter(Q(tipoacta__graduado__inscripcion__persona__nombres__icontains=search) |
                                                     Q(tipoacta__graduado__inscripcion__persona__apellido1__icontains=search) |
                                                     Q(tipoacta__graduado__inscripcion__persona__apellido2__icontains=search) |
                                                     Q(tipoacta__graduado__inscripcion__persona__cedula__icontains=search) |
                                                     Q(tipoacta__graduado__inscripcion__persona__pasaporte__icontains=search))
                        else:
                            listado = listado.filter(Q(tipoacta__graduado__inscripcion__persona__apellido1__icontains=ss[0]) &
                                                     Q(tipoacta__graduado__inscripcion__persona__apellido2__icontains=ss[1]))
                    idc = 0
                    if 'idc' in request.GET:
                        idc = int(request.GET['idc'])
                        url_vars += f"&idc={idc}"
                        if idc > 0:
                            listado = listado.filter(tipoacta__graduado__inscripcion__carrera__id=idc)
                    fechini = None
                    fechfin = None
                    if 'fechini' in request.GET and 'fechfin' in request.GET:
                        fechini = request.GET['fechini']
                        fechfin = request.GET['fechfin']
                        if fechini and fechfin:
                            url_vars += f"&fechini={fechini}"
                            url_vars += f"&fechfin={fechfin}"
                            listado = listado.filter(tipoacta__graduado__fechagraduado__gte=fechini, tipoacta__graduado__fechagraduado__lte=fechfin)
                    numerofilas = 25
                    paging = MiPaginador(listado, numerofilas)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                            if p == 1:
                                numerofilasguiente = numerofilas
                            else:
                                numerofilasguiente = numerofilas * (p - 1)
                        else:
                            p = paginasesion
                            if p == 1:
                                numerofilasguiente = numerofilas
                            else:
                                numerofilasguiente = numerofilas * (p - 1)
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['numeropagina'] = p
                    data['numerofilasguiente'] = numerofilasguiente
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['idc'] = idc
                    data['fechini'] = fechini
                    data['fechfin'] = fechfin
                    data['search'] = search if search else ""
                    data['url_vars'] = url_vars
                    data['listado'] = page.object_list
                    request.session['viewactivoth'] = ['pendienteactagrado', 'listadofirmadasactasgrado']
                    return render(request, "graduados/listadoalumnosfirmar.html", data)
                except Exception as ex:
                    pass

            elif action == 'listadoactasconsolidada':
                try:
                    url_vars, search, = f'&action={action}', request.GET.get('s', '').strip()
                    data['title'] = u'Listado de actas consolidadas'
                    data['horaactual'] = horaactual
                    listado = persona.tipoactafirma_set.filter(tipoacta__tipo=6, firmado=False, turnofirmar=True, status=True)
                    data['carreras'] = Carrera.objects.filter(pk__in=listado.values_list('tipoacta__graduado__inscripcion__carrera_id', flat=True).distinct(), status=True).order_by('nombre')
                    if 's' in request.GET:
                        search = request.GET['s']
                        url_vars += f"&s={search}"
                        ss = search.split(' ')
                        if len(ss) == 1:
                            listado = listado.filter(Q(tipoacta__graduado__inscripcion__persona__nombres__icontains=search) |
                                                     Q(tipoacta__graduado__inscripcion__persona__apellido1__icontains=search) |
                                                     Q(tipoacta__graduado__inscripcion__persona__apellido2__icontains=search) |
                                                     Q(tipoacta__graduado__inscripcion__persona__cedula__icontains=search) |
                                                     Q(tipoacta__graduado__inscripcion__persona__pasaporte__icontains=search))
                        else:
                            listado = listado.filter(Q(tipoacta__graduado__inscripcion__persona__apellido1__icontains=ss[0]) &
                                                     Q(tipoacta__graduado__inscripcion__persona__apellido2__icontains=ss[1]))
                    idc = 0
                    if 'idc' in request.GET:
                        idc = int(request.GET['idc'])
                        url_vars += f"&idc={idc}"
                        if idc > 0:
                            listado = listado.filter(tipoacta__graduado__inscripcion__carrera__id=idc)
                    fechini = None
                    fechfin = None
                    if 'fechini' in request.GET and 'fechfin' in request.GET:
                        fechini = request.GET['fechini']
                        fechfin = request.GET['fechfin']
                        if fechini and fechfin:
                            url_vars += f"&fechini={fechini}"
                            url_vars += f"&fechfin={fechfin}"
                            listado = listado.filter(tipoacta__graduado__fechagraduado__gte=fechini, tipoacta__graduado__fechagraduado__lte=fechfin)
                    numerofilas = 25
                    paging = MiPaginador(listado, numerofilas)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                            if p == 1:
                                numerofilasguiente = numerofilas
                            else:
                                numerofilasguiente = numerofilas * (p - 1)
                        else:
                            p = paginasesion
                            if p == 1:
                                numerofilasguiente = numerofilas
                            else:
                                numerofilasguiente = numerofilas * (p - 1)
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['numeropagina'] = p
                    data['numerofilasguiente'] = numerofilasguiente
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['idc'] = idc
                    data['fechini'] = fechini
                    data['fechfin'] = fechfin
                    data['search'] = search if search else ""
                    data['url_vars'] = url_vars
                    data['listado'] = page.object_list
                    request.session['viewactivoth'] = ['pendienteactaconsolidada', 'listadoactasconsolidada']
                    return render(request, "graduados/listadoactasconsolidada.html", data)
                except Exception as ex:
                    pass

            elif action == 'listadofirmadaactasconsolidada':
                try:
                    puede_eliminar_actas_firmadas = variable_valor('PUEDE_ELIMINAR_ACTAS_FIRMADAS')
                    url_vars, search, = f'&action={action}', request.GET.get('s', '').strip()
                    data['title'] = u'Actas consolidadas firmadas'
                    data['horaactual'] = horaactual
                    deleteactascons = True
                    if user.is_superuser:
                        deleteactascons = False
                        listado = TipoActaFirma.objects.filter(tipoacta__tipo=6, firmado=True, status=True, orden=1).order_by('-tipoacta__graduado__fechagraduado','tipoacta__graduado__inscripcion__persona__apellido1')
                    else:
                        listado = persona.tipoactafirma_set.filter(tipoacta__tipo=6, firmado=True, status=True).order_by('-tipoacta__graduado__fechagraduado','tipoacta__graduado__inscripcion__persona__apellido1')
                    data['deleteactascons'] = deleteactascons
                    data['puede_eliminar_actas_firmadas'] = puede_eliminar_actas_firmadas
                    data['carreras'] = Carrera.objects.filter(pk__in=listado.values_list('tipoacta__graduado__inscripcion__carrera_id', flat=True).distinct(), status=True).order_by('nombre')
                    if 's' in request.GET:
                        search = request.GET['s']
                        url_vars += f"&s={search}"
                        ss = search.split(' ')
                        if len(ss) == 1:
                            listado = listado.filter(Q(tipoacta__graduado__inscripcion__persona__nombres__icontains=search) |
                                                     Q(tipoacta__graduado__inscripcion__persona__apellido1__icontains=search) |
                                                     Q(tipoacta__graduado__inscripcion__persona__apellido2__icontains=search) |
                                                     Q(tipoacta__graduado__inscripcion__persona__cedula__icontains=search) |
                                                     Q(tipoacta__graduado__inscripcion__persona__pasaporte__icontains=search),
                                                     status=True)
                        else:
                            listado = listado.filter(Q(tipoacta__graduado__inscripcion__persona__apellido1__icontains=ss[0]) &
                                                     Q(tipoacta__graduado__inscripcion__persona__apellido2__icontains=ss[1]),
                                                     status=True)
                    idc = 0
                    if 'idc' in request.GET:
                        idc = int(request.GET['idc'])
                        url_vars += f"&idc={idc}"
                        if idc > 0:
                            listado = listado.filter(tipoacta__graduado__inscripcion__carrera__id=idc)
                    fechini = None
                    fechfin = None
                    if 'fechini' in request.GET and 'fechfin' in request.GET:
                        fechini = request.GET['fechini']
                        fechfin = request.GET['fechfin']
                        if fechini and fechfin:
                            url_vars += f"&fechini={fechini}"
                            url_vars += f"&fechfin={fechfin}"
                            listado = listado.filter(tipoacta__graduado__fechagraduado__gte=fechini, tipoacta__graduado__fechagraduado__lte=fechfin)
                    numerofilas = 25
                    paging = MiPaginador(listado, numerofilas)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                            if p == 1:
                                numerofilasguiente = numerofilas
                            else:
                                numerofilasguiente = numerofilas * (p - 1)
                        else:
                            p = paginasesion
                            if p == 1:
                                numerofilasguiente = numerofilas
                            else:
                                numerofilasguiente = numerofilas * (p - 1)
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['numeropagina'] = p
                    data['numerofilasguiente'] = numerofilasguiente
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['idc'] = idc
                    data['fechini'] = fechini
                    data['fechfin'] = fechfin
                    data['search'] = search if search else ""
                    data['url_vars'] = url_vars
                    data['listado'] = page.object_list
                    request.session['viewactivoth'] = ['pendienteactaconsolidada', 'listadofirmadaactasconsolidada']
                    return render(request, "graduados/listadoactasconsolidada.html", data)
                except Exception as ex:
                    pass

            elif action == 'listadofirmatotalconsolidada':
                try:
                    url_vars, search, = f'&action={action}', request.GET.get('s', '').strip()
                    data['title'] = u'Listado de alumnos'
                    data['horaactual'] = horaactual
                    deleteactascons = True
                    if user.is_superuser:
                        deleteactascons = False
                        listado = TipoActaFirma.objects.filter(tipoacta__tipo=6, firmado=True, tipoacta__actafirmada=True, status=True, orden=1).order_by('-tipoacta__graduado__fechagraduado','tipoacta__graduado__inscripcion__persona__apellido1','tipoacta__graduado__inscripcion__persona__apellido2')
                    else:
                        listado = persona.tipoactafirma_set.filter(tipoacta__tipo=6, firmado=True, tipoacta__actafirmada=True, status=True, orden=1).order_by('-tipoacta__graduado__fechagraduado','tipoacta__graduado__inscripcion__persona__apellido1','tipoacta__graduado__inscripcion__persona__apellido2')
                    data['deleteactascons'] = deleteactascons
                    data['carreras'] = Carrera.objects.filter(pk__in=listado.values_list('tipoacta__graduado__inscripcion__carrera_id', flat=True).distinct(), status=True).order_by('nombre')
                    if 's' in request.GET:
                        search = request.GET['s']
                        url_vars += f"&s={search}"
                        ss = search.split(' ')
                        if len(ss) == 1:
                            listado = listado.filter(Q(tipoacta__graduado__inscripcion__persona__nombres__icontains=search) |
                                                     Q(tipoacta__graduado__inscripcion__persona__apellido1__icontains=search) |
                                                     Q(tipoacta__graduado__inscripcion__persona__apellido2__icontains=search) |
                                                     Q(tipoacta__graduado__inscripcion__persona__cedula__icontains=search) |
                                                     Q(tipoacta__graduado__inscripcion__persona__pasaporte__icontains=search),
                                                     status=True)
                        else:
                            listado = listado.filter(Q(tipoacta__graduado__inscripcion__persona__apellido1__icontains=ss[0]) &
                                                     Q(tipoacta__graduado__inscripcion__persona__apellido2__icontains=ss[1]),
                                                     status=True)
                    idc = 0
                    if 'idc' in request.GET:
                        idc = int(request.GET['idc'])
                        url_vars += f"&idc={idc}"
                        if idc > 0:
                            listado = listado.filter(tipoacta__graduado__inscripcion__carrera__id=idc)
                    fechini = None
                    fechfin = None
                    if 'fechini' in request.GET and 'fechfin' in request.GET:
                        fechini = request.GET['fechini']
                        fechfin = request.GET['fechfin']
                        if fechini and fechfin:
                            url_vars += f"&fechini={fechini}"
                            url_vars += f"&fechfin={fechfin}"
                            listado = listado.filter(tipoacta__graduado__fechagraduado__gte=fechini, tipoacta__graduado__fechagraduado__lte=fechfin)
                    numerofilas = 25
                    paging = MiPaginador(listado, numerofilas)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                            if p == 1:
                                numerofilasguiente = numerofilas
                            else:
                                numerofilasguiente = numerofilas * (p - 1)
                        else:
                            p = paginasesion
                            if p == 1:
                                numerofilasguiente = numerofilas
                            else:
                                numerofilasguiente = numerofilas * (p - 1)
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['numeropagina'] = p
                    data['numerofilasguiente'] = numerofilasguiente
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['idc'] = idc
                    data['fechini'] = fechini
                    data['fechfin'] = fechfin
                    data['search'] = search if search else ""
                    data['url_vars'] = url_vars
                    data['listado'] = page.object_list
                    request.session['viewactivoth'] = ['documentosalumnos', 'listadofirmatotalconsolidada']
                    return render(request, "graduados/listadosubirtitulos.html", data)
                except Exception as ex:
                    pass

            elif action == 'listadorevisiondocumentos':
                try:
                    url_vars, search, = f'&action={action}', request.GET.get('s', '').strip()
                    data['title'] = u'Revisión de documentos'
                    data['horaactual'] = horaactual
                    listado = Graduado.objects.filter(inscripcion__carrera__in=miscarreras, fechagraduado__isnull=False, status=True).order_by('-fechagraduado','inscripcion__persona__apellido1','inscripcion__persona__apellido2')
                    data['carreras'] = Carrera.objects.filter(pk__in=listado.values('inscripcion__carrera_id').distinct(), status=True).order_by('nombre')
                    if 's' in request.GET:
                        search = request.GET['s']
                        url_vars += f"&s={search}"
                        ss = search.split(' ')
                        if len(ss) == 1:
                            listado = listado.filter(Q(inscripcion__persona__nombres__icontains=search) |
                                                     Q(inscripcion__persona__apellido1__icontains=search) |
                                                     Q(inscripcion__persona__apellido2__icontains=search) |
                                                     Q(inscripcion__persona__cedula__icontains=search) |
                                                     Q(inscripcion__persona__pasaporte__icontains=search),
                                                     status=True)
                        else:
                            listado = listado.filter(Q(inscripcion__persona__apellido1__icontains=ss[0]) &
                                                     Q(inscripcion__persona__apellido2__icontains=ss[1]),
                                                     status=True)
                    idc = 0
                    if 'idc' in request.GET:
                        idc = int(request.GET['idc'])
                        url_vars += f"&idc={idc}"
                        if idc > 0:
                            listado = listado.filter(inscripcion__carrera__id=idc)

                    fechini = None
                    fechfin = None
                    if 'fechini' in request.GET and 'fechfin' in request.GET:
                        fechini = request.GET['fechini']
                        fechfin = request.GET['fechfin']
                        if fechini and fechfin:
                            url_vars += f"&fechini={fechini}"
                            url_vars += f"&fechfin={fechfin}"
                            listado = listado.filter(fechagraduado__gte=fechini, fechagraduado__lte=fechfin)

                    numerofilas = 25
                    paging = MiPaginador(listado, numerofilas)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                            if p == 1:
                                numerofilasguiente = numerofilas
                            else:
                                numerofilasguiente = numerofilas * (p - 1)
                        else:
                            p = paginasesion
                            if p == 1:
                                numerofilasguiente = numerofilas
                            else:
                                numerofilasguiente = numerofilas * (p - 1)
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['numeropagina'] = p
                    data['numerofilasguiente'] = numerofilasguiente
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['idc'] = idc
                    data['fechini'] = fechini
                    data['fechfin'] = fechfin
                    data['search'] = search if search else ""
                    data['url_vars'] = url_vars
                    data['listado'] = page.object_list
                    request.session['viewactivoth'] = ['documentosalumnos', 'listadorevisiondocumentos']
                    return render(request, "graduados/revisiondocumentos.html", data)
                except Exception as ex:
                    pass

            elif action == 'detallefirmasactagrado':
                try:
                    data = {}
                    data['tipoacta'] = tipoacta = TipoActa.objects.get(pk=int(request.GET['id']))
                    data['listadofirmas'] = tipoacta.tipoactafirma_set.filter(status=True).order_by('orden')
                    template = get_template("graduados/detallefirmas.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": True, 'data': json_content})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            elif action == 'generaactagrado':
                try:
                    data = {}
                    data['horaactual'] = horaactual
                    tipoacta = TipoActa.objects.get(pk=int(request.GET['id']))
                    generaactagradofe = actagradocomplexivofirma(tipoacta.graduado.id)
                    if generaactagradofe:
                        tipoacupdate = TipoActa.objects.filter(tipo=5, graduado=tipoacta.graduado)[0]
                        tipoacupdate.archivo = 'qrcode/actatitulacion/fe_actagrado_' + str(tipoacta.graduado.id) + '.pdf'
                        tipoacupdate.save(request)
                        random_number = random.randint(1, 1000000)
                        data['archivo'] = tipoacupdate.archivo.url
                        archivo = f"{tipoacupdate.archivo.url}?cache={random_number}"
                        data['url_archivo'] = '{}{}'.format(dominio_sistema, archivo)
                        template = get_template("graduados/viewactagrado.html")
                        json_content = template.render(data)
                        return JsonResponse({"result": True, 'data': json_content})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            elif action == 'generaactaconsolidada':
                try:
                    data = {}
                    data['horaactual'] = horaactual
                    tipoacta = TipoActa.objects.get(pk=int(request.GET['id']))
                    generaactagradofe = actaconsolidadafirma(tipoacta.graduado.id)
                    if generaactagradofe:
                        tipoacupdate = TipoActa.objects.filter(tipo=6, graduado=tipoacta.graduado)[0]
                        tipoacupdate.archivo = 'qrcode/actatitulacion/fe_actaconsolidada_' + str(tipoacta.graduado.id) + '.pdf'
                        tipoacupdate.save(request)
                        # actualiza usuarios que firman
                        listadopersonafirman = tipoacupdate.tipoactafirma_set.filter(status=True)
                        for lpersonas in listadopersonafirman:
                            if lpersonas.orden == 2:
                                for dirfac in tipoacta.graduado.directoresfacultad.all():
                                    dirfacultad = dirfac
                                lpersonas.persona = dirfacultad
                                lpersonas.save(request)
                            if lpersonas.orden == 3:
                                lpersonas.persona_id = tipoacta.graduado.decano_id
                                lpersonas.save(request)
                            if lpersonas.orden == 4:
                                lpersonas.persona_id = tipoacta.graduado.secretariageneral_id
                                lpersonas.save(request)

                        random_number = random.randint(1, 1000000)
                        data['archivo'] = tipoacupdate.archivo.url
                        archivo = f"{tipoacupdate.archivo.url}?cache={random_number}"
                        data['url_archivo'] = '{}{}'.format(dominio_sistema, archivo)
                        template = get_template("graduados/viewactagrado.html")
                        json_content = template.render(data)
                        return JsonResponse({"result": True, 'data': json_content})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            elif action == 'consultaracta':
                try:
                    nombre = request.GET['nombre']
                    data['actas'] = actas = TipoActa.objects.filter(status=True, graduado=int(request.GET['id']), nombre=nombre)
                    if actas:
                        archivo = actas[0].archivo
                        return JsonResponse({"result": "ok", "ruta": str(archivo)})
                    else:
                        return JsonResponse({"result":"bad"})
                except Exception as ex:
                    pass

            elif action == 'modalsubirarchivo':
                try:
                    template = get_template("graduados/addacta.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'adicionaractagraduado':
                try:
                    form = ActaGraduadosForm()
                    form.fields['acta'].initial = data['tipoacta'] = int(request.GET['tipoacta'])
                    form.sololectura()
                    data['form2'] = form
                    data['action'] = 'adicionaractagraduado'
                    data['graduado'] = int(request.GET['id'])
                    template = get_template('graduados/modal/formregistroactagraduados.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'reporte_generar_masivo':
                try:
                    limiteinsignias = 500
                    graduados = Graduado.objects.filter(Q(status=True, inscripcion__carrera__coordinacion__id=7),
                                                        Q(Q(rutapdftitulo__isnull=True) | Q(urlhtmltitulo__isnull=True) | Q(namehtmltitulo__isnull=True)),
                                                        fecharefrendacion__isnull=False, fecharefrendacion__year__gte=2022).order_by('-id')
                    # , fecharefrendacion__date__lte = datetime.now().date()
                    if graduados:
                        # pruebas local
                        if IS_DEBUG and len(graduados) > 2:
                            graduados = graduados[:2]

                        if len(graduados) > limiteinsignias:
                            graduados = graduados[:limiteinsignias]
                        resultados_errores = generartituloeinsigniaposgrado(request, graduados, data, procesomasivo=True, IS_DEBUG=IS_DEBUG)
                        if len(resultados_errores) > 0:
                            return JsonResponse({"result": "bad", "mensaje": u"Problemas al generar los Títulos, recibirá una notificación."})
                        else:
                            return JsonResponse({"result": "ok", "mensaje": u"Títulos masivos generados exitosamente, recibirá una notificación."})
                    else:
                        return JsonResponse({"result": "bad", "mensaje": u"No hay nuevos registros de graduados para generar titulos, los anteriores ya fueron generados."})
                except Exception as ex:
                    messages.error(request, f'Error {ex} on line {sys.exc_info()[-1].tb_lineno}')
                    return JsonResponse({"result": "bad", "mensaje": f"Error al obtener los datos {str(ex)}. Error on line {sys.exc_info()[-1].tb_lineno}"})

            elif action == 'reporte_generar_masivo_webpush':
                # Generar titulo/insignia masivo web push
                try:
                    graduados = Graduado.objects.filter(Q(status=True, inscripcion__carrera__coordinacion__id=7),
                                                        Q(Q(rutapdftitulo__isnull=True) | Q(
                                                            urlhtmltitulo__isnull=True) | Q(
                                                            namehtmltitulo__isnull=True)),
                                                        fecharefrendacion__isnull=False,
                                                        fecharefrendacion__year__gte=2022).order_by('-id')
                    if graduados:
                        # pruebas local
                        if IS_DEBUG and len(graduados) > 3:
                            graduados = graduados[:3]

                        # notificacion
                        noti = Notificacion(cuerpo='Generación de títulos/insignia posgrado masivo en progreso...',
                                            titulo='Generación de títulos/insignia posgrado masivo',
                                            destinatario=persona,
                                            # url=f"{'https://sga.unemi.edu.ec/graduados' if not IS_DEBUG else 'http://127.0.0.1:8000'}/graduados",
                                            url=f"/graduados",
                                            prioridad=1,
                                            app_label='SGA',
                                            fecha_hora_visible=datetime.now() + timedelta(days=1),
                                            tipo=2,
                                            en_proceso=True)
                        noti.save(request)
                        # generacion
                        reporte_tituloinsigniamasivo_background(request=request,
                                                                notiid=noti.pk,
                                                                graduados = graduados,
                                                                data = data,
                                                                IS_DEBUG=IS_DEBUG
                                                                ).start()
                        # resultado
                        return JsonResponse({"result": "ok",
                                             "mensaje": u'La generación de títulos se está realizando. Verifique su apartado "Ver mis notificaciones" después de unos minutos.',
                                             "btn_notificaciones": traerNotificaciones(request, data, persona)})

                    else:
                        return JsonResponse({"result": "bad",
                                             "mensaje": u"No hay nuevos registros de graduados para generar titulos, los anteriores ya fueron generados."})
                except Exception as ex:
                    messages.error(request, f'Error {ex} on line {sys.exc_info()[-1].tb_lineno}')
                    return JsonResponse({"result": "bad",
                                         "mensaje": f"Error al obtener los datos {str(ex)}. Error on line {sys.exc_info()[-1].tb_lineno}"})


            elif action == 'notificar_insignia_masivo':
                try:
                    graduados = Graduado.objects.filter(status=True, inscripcion__carrera__coordinacion__id=7,
                                                        rutapdftitulo__isnull=False, urlhtmltitulo__isnull=False,
                                                        namehtmltitulo__isnull=False, estadonotificacion=2)
                    if graduados:
                        resulnotificacionerrores = solonotificarcorreoinsigniaposgrado(request, graduados,
                                                                                       procesomasivo=True,
                                                                                       IS_DEBUG=IS_DEBUG)
                        if len(resulnotificacionerrores) > 0:
                            return JsonResponse(
                                {"result": "bad", "mensaje": u"Problemas al notificar de forma masiva al email."})
                        else:
                            return JsonResponse({"result": "ok", "mensaje": u"Emails enviados correctamente."})
                    else:
                        return JsonResponse({"result": "bad",
                                             "mensaje": u"No hay insignias/títulos pendientes de notificar. Por favor, genere primero las insignias y títulos."})
                except Exception as ex:
                    return JsonResponse(
                        {"result": "bad", "mensaje": u"Problemas al enviar las notificaciones al email. %s" % ex})

            elif action == 'importarrefrendacionpos':
                try:
                    data['id_persona'] = request.GET['id']
                    data['action'] = 'importarrefrendacionpos'
                    template = get_template("graduados/modal/importarrefrendacionpos.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    transaction.set_rollback(True)
                    messages.error(request, str(ex))
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            elif action == 'firmarinformesmasivo':
                try:
                    ids = None
                    if 'ids' in request.GET:
                        ids = request.GET['ids']
                    leadsselect = ids
                    data['listadoseleccion'] = leadsselect
                    data['accionfirma'] = request.GET['accionfirma']
                    template = get_template("adm_criteriosactividadesdocente/firmarinformesmasivo.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": True, 'html': json_content})
                except Exception as ex:
                    mensaje = 'Intentelo más tarde'
                    return JsonResponse({"result": False, "mensaje": mensaje})

            return HttpResponseRedirect(request.path)

        else:
            data['title'] = u'Listado de alumnos graduados'

            search = None
            ids = None
            inscripcionid = None
            graduados = Graduado.objects.filter(status=True,inscripcion__carrera__in=miscarreras).order_by('inscripcion__carrera__nombre', 'inscripcion__persona__apellido1', 'inscripcion__persona__apellido2')
            if 'id' in request.GET:
                ids = request.GET['id']
                graduados = graduados.filter(pk=ids)
            if 'inscripcionid' in request.GET:
                inscripcionid = request.GET['inscripcionid']
                graduados = graduados.filter(inscripcion__id=inscripcionid)
            if 's' in request.GET:
                search = request.GET['s']
                ss = search.split(' ')
                if len(ss) == 1:
                    graduados = graduados.filter(Q(inscripcion__persona__nombres__icontains=search) |
                                                 Q(inscripcion__persona__apellido1__icontains=search) |
                                                 Q(inscripcion__persona__apellido2__icontains=search) |
                                                 Q(inscripcion__persona__cedula__icontains=search) |
                                                 Q(inscripcion__persona__pasaporte__icontains=search))
                else:
                    graduados = graduados.filter(Q(inscripcion__persona__apellido1__icontains=ss[0]) &
                                                 Q(inscripcion__persona__apellido2__icontains=ss[1]))
            idc = 0
            if 'idc' in request.GET:
                idc = int(request.GET['idc'])
                if idc > 0:
                    graduados = graduados.filter(inscripcion__carrera__id=idc)
            idg = 0
            if 'idg' in request.GET:
                idg = int(request.GET['idg'])
                if idg > 0:
                    graduados = graduados.filter(inscripcion__persona__sexo__id=idg)
            ida = 0
            if 'ida' in request.GET and request.GET['ida']!='undefined':
                ida = int(request.GET['ida'])
                if ida > 0:
                    graduados = graduados.filter(fechagraduado__year=ida)

            paging = MiPaginador(graduados, 30)
            p = 1
            try:
                paginasesion = 1
                if 'paginador' in request.session:
                    paginasesion = int(request.session['paginador'])
                if 'page' in request.GET:
                    p = int(request.GET['page'])
                else:
                    p = paginasesion
                try:
                    page = paging.page(p)
                except:
                    p = 1
                page = paging.page(p)
            except:
                page = paging.page(p)
            request.session['paginador'] = p
            data['paging'] = paging
            data['page'] = page
            data['rangospaging'] = paging.rangos_paginado(p)
            data['graduados'] = page.object_list
            data['search'] = search if search else ""
            data['ids'] = ids if ids else ""
            data['inscripcionid'] = inscripcionid if inscripcionid else ""
            data['idg'] = idg
            data['ida'] = ida
            data['idc'] = idc
            data['version'] = datetime.now().strftime('%Y%m%d_%H%M%S%f')
            # data['carreras'] = Carrera.objects.filter(status=True, coordinacion__excluir=False).order_by('nombre')
            data['carreras'] = miscarreras.filter(status=True, coordinacion__excluir=False).order_by('nombre')
            data['sexo'] = Sexo.objects.filter(status=True)
            data['anios'] = Graduado.objects.filter(status=True, inscripcion__carrera__in=miscarreras, fechagraduado__isnull=False).annotate(Year=ExtractYear('fechagraduado')).values_list('Year', flat=True).order_by('Year').distinct()
            form = ActaFacultadForm()
            form.adicionar()
            data['form2'] = form
            puede_ver_opcion = False
            if puede_realizar_accion_afirmativo(request, 'sga.puede_descargar_db_backup'):
                puede_ver_opcion = True
            data['puede_ver_opcion'] = puede_ver_opcion
            return render(request, "graduados/view.html", data)


def generartituloeinsigniaposgrado(request, graduados, data, vistaprevia = False, procesomasivo=False, IS_DEBUG=False):
    dominio_sistema = 'http://127.0.0.1:8000'
    if not IS_DEBUG:
        dominio_sistema = 'https://sga.unemi.edu.ec'
    data["DOMINIO_DEL_SISTEMA"] = dominio_sistema

    lista_correctos = []
    lista_errores = []
    for graduado in graduados:
        # with transaction.atomic():
        try:
            persona_cargo_tercernivel = None
            cargo = None
            tamano = 0
            firmauno = None
            if DistributivoPersona.objects.filter(denominacionpuesto=113, estadopuesto__id=1, status=True).exists():
                firmauno = DistributivoPersona.objects.filter(denominacionpuesto=113, estadopuesto__id=1, status=True)[0]
            data['firmauno'] = firmauno
            firmatres = None
            # Cambio de Secretaria General Velasco Neira Stefania desde el 25 de oct 2022 hasta el 12 de feb 2023 subrogante y desde el 14 de feb 2023 como titular, 13 feb 2023 Diana Pincay
            firmatres_tipocargo = None

            firmatres = PersonaDepartamentoFirmas.objects.filter(status=True, departamento_id = 43,activo=True,actualidad=True).first()

            ePersonaDepartamentoFirmas = PersonaDepartamentoFirmas.objects.filter(status=True, departamento_id = 43)
            for firma in ePersonaDepartamentoFirmas:
                if firma.fechafin is not None and firma.fechainicio is not None:
                    if graduado.fecharefrendacion:
                        if graduado.fecharefrendacion <= firma.fechafin and graduado.fecharefrendacion >= firma.fechainicio:
                            firmatres = firma

            # Tercera firma
            data['firmatres_rubrica'] = FirmaPersona.objects.filter(status=True, persona=firmatres.personadepartamento, tipofirma=1).first().firma
            data['firmatres_tipocargo'] = firmatres.tiposubrogante.abreviatura
            # Tercera firma
            data['firmatres'] = firmatres
            firma_departamento = PersonaDepartamentoFirmas.objects.get(tipopersonadepartamento_id=1,
                                                                       departamentofirma_id=1, status=True,
                                                                       actualidad=True)
            listafirmaspersonadepartamento = PersonaDepartamentoFirmas.objects.filter(tipopersonadepartamento_id=1,
                                                                                      departamentofirma_id=1,
                                                                                      status=True)
            for firma in listafirmaspersonadepartamento:
                if firma.fechafin is not None and firma.fechainicio is not None:
                    if graduado.fecharefrendacion <= firma.fechafin and graduado.fecharefrendacion >= firma.fechainicio:
                        firma_departamento = firma
            data['firmadirector'] = firma_departamento
            data['imgfirmadirector'] = firma_departamento.personadepartamento.firmapersona_set.filter(status=True, tipofirma=1).order_by('-tipofirma').first()
            # fin firma vicerrector/directo
            if PersonaDepartamentoFirmas.objects.filter(actualidad=True, status=True).exists():
                firmaizquierda = PersonaDepartamentoFirmas.objects.get(actualidad=True, status=True,
                                                                       tipopersonadepartamento_id=2,
                                                                       departamentofirma_id=1)
            # if PersonaDepartamentoFirmas.objects.values('id').filter(status=True,
            #                                                          fechafin__gte=evento.fechafin,
            #                                                          fechainicio__lte=evento.fechafin,
            #                                                          tipopersonadepartamento_id=2,
            #                                                          departamentofirma_id=1).exists():
            #     firmaizquierda = PersonaDepartamentoFirmas.objects.get(status=True,
            #                                                            fechafin__gte=evento.fechafin,
            #                                                            fechainicio__lte=evento.fechafin,
            #                                                            tipopersonadepartamento_id=2,
            #                                                            departamentofirma_id=1)
            #
            data['firmaizquierda'] = firmaizquierda
            data['firmaimgizq'] = FirmaPersona.objects.filter(status=True,
                                                              persona=firmaizquierda.personadepartamento, tipofirma=1).last()
            # if evento.envionotaemail:
            #     data['nota'] = evento.instructor_principal().extaer_notatotal(graduado.id)
            # if DistributivoPersona.objects.filter(persona_id=persona, estadopuesto__id=PUESTO_ACTIVO_ID,
            #                                       status=True).exists():
            #     cargo = \
            #         DistributivoPersona.objects.filter(persona_id=persona,
            #                                            estadopuesto__id=PUESTO_ACTIVO_ID,
            #                                            status=True)[0]
            data['persona_cargo'] = cargo
            # data['persona_cargo_titulo'] = titulo = persona.titulacion_principal_senescyt_registro()
            # if not titulo == '':
            #     persona_cargo_tercernivel = \
            #         persona.titulacion_set.filter(titulo__nivel=3).order_by('-fechaobtencion')[
            #             0] if titulo.titulo.nivel_id == 4 else None
            data['persona_cargo_tercernivel'] = persona_cargo_tercernivel
            data['graduado'] = graduado
            # data['title'] = ''
            data['fecha'] = None
            fechagraduado = graduado.fecharefrendacion if graduado.fecharefrendacion else None
            if fechagraduado:
                mes = ["enero", "febrero", "marzo", "abril", "mayo", "junio", "julio", "agosto", "septiembre",
                       "octubre", "noviembre", "diciembre"]
                if fechagraduado.day > 1:
                    cadena = u"a los %s días" % (fechagraduado.day)
                else:
                    cadena = u"al primer día"
                data['fecha'] = u"San Francisco de Milagro, " + cadena + " del mes de %s de %s." % (
                    str(mes[fechagraduado.month - 1]), fechagraduado.year)
            data['fecha_graduado_insignia'] = fechagraduado

            # data['listado_contenido'] = listado = evento.contenido.split("\n") if evento.contenido else []
            # if evento.objetivo.__len__() < 290:
            #     if listado.__len__() < 21:
            #         tamano = 120
            #     elif listado.__len__() < 35:
            #         tamano = 100
            #     elif listado.__len__() < 41:
            #         tamano = 70
            data['controlar_bajada_logo'] = tamano
            qrname = 'qr_titulo_' + str(graduado.id)
            # folder = SITE_STORAGE + 'media/qrcode/evaluaciondocente/'
            folder = os.path.join(os.path.join(SITE_STORAGE, 'media', 'qrcode', 'titulos', 'qr'))
            directory = os.path.join(os.path.join(SITE_STORAGE, 'media', 'qrcode', 'titulos'))
            # folder = os.path.join(SITE_STORAGE, 'media', 'qrcode', 'evaluaciondocente')
            rutapdf = folder + qrname + '.pdf'
            rutaimg = folder + qrname + '.png'
            try:
                os.stat(directory)
            except:
                os.mkdir(directory)
            if os.path.isfile(rutapdf):
                os.remove(rutaimg)
                os.remove(rutapdf)
            # generar nombre html y url html
            if not graduado.namehtmltitulo:
                htmlname = "%s%s" % (uuid.uuid4().hex, '.html')
            else:
                htmlname = graduado.namehtmltitulo
            urlname = "/media/qrcode/titulos/%s" % htmlname
            rutahtml = SITE_STORAGE + urlname
            if os.path.isfile(rutahtml):
                os.remove(rutahtml)
            # generar nombre html y url html
            data['version'] = version = datetime.now().strftime('%Y%m%d_%H%M%S%f')
            # url = pyqrcode.create('https://sga.unemi.edu.ec/media/qrcode/titulos/' + qrname + '.pdf')
            url = pyqrcode.create(f'{dominio_sistema}/media/qrcode/titulos/{htmlname}?v={version}')
            # url = pyqrcode.create(dominio_sistema + '/media/qrcode/titulos/' + htmlname)
            imageqr = url.png(folder + qrname + '.png', 16, '#000000')
            data['qrname'] = 'qr' + qrname
            data['urlhtmlinsignia'] = urlhtmlinsignia = dominio_sistema + urlname
            # data['urlhtmlinsigniaversion'] = f'{urlhtmlinsignia}?v={version}'
            data['posgrado'] = u'FACULTAD DE POSGRADO'
            valida = conviert_html_to_pdfsaveqrtitulo(
                'graduados/titulo_formatonuevo_pdf.html',
                {'pagesize': 'A4', 'data': data},
                qrname + '.pdf',
                vistaprevia
            )
            if valida:
                # elimino codigo qr despues de pegarlo en el titulo titulo_formatonuevo_pdf.html
                if os.path.isfile(rutaimg):
                    os.remove(rutaimg)
                # generar portada del certificado
                portada = convertir_certificadopdf_a_jpg(qrname, SITE_STORAGE, dominio_sistema, data)
                graduado.rutapdftitulo = 'qrcode/titulos/' + qrname + '.pdf'
                # graduado.emailnotificado = True
                # graduado.fecha_emailnotifica = datetime.now().date()
                # graduado.persona_emailnotifica = persona
                # graduado.save(request)
                data['rutapdf'] = '/media/{}'.format(graduado.rutapdftitulo) # ojo
                #
                data['idinsignia'] = htmlname[0:len(htmlname) - 5]

                # consultar record academico
                data['inscripcion'] = inscripcion = graduado.inscripcion
                data['records'] = inscripcion.recordacademico_set.filter(status=True).order_by(
                    'asignaturamalla__nivelmalla', 'asignatura', 'fecha')
                data['total_creditos'] = inscripcion.total_creditos()
                data['total_creditos_malla'] = inscripcion.total_creditos_malla()
                data['total_creditos_modulos'] = inscripcion.total_creditos_modulos()
                data['total_creditos_otros'] = inscripcion.total_creditos_otros()
                data['total_horas'] = inscripcion.total_horas()
                data['promedio'] = inscripcion.promedio_record()
                # data['aprobadas'] = inscripcion.recordacademico_set.filter(aprobada=True, valida=True).count()
                # data['reprobadas'] = inscripcion.recordacademico_set.filter(aprobada=False, valida=True).count()
                # data['reporte_0'] = obtener_reporte("record_alumno")
                # consultar record academico
                # crear html de titulo valido en la media  y guardar url en base
                a = render(request, "graduados/titulovalido.html",
                           {"data": data, 'institucion': 'UNIVERSIDAD ESTATAL DE MILAGRO', "remotenameaddr": 'sga.unemi.edu.ec'})
                with open(SITE_STORAGE + urlname, "wb") as f:
                    f.write(a.content)
                f.close()
                # elimino portada de titulo despues de añadirla en el insignia titulovalido.html
                # if os.path.isfile(data['ruta_jpg']):
                #     os.remove(data['ruta_jpg'])
                if not vistaprevia:
                    graduado.namehtmltitulo = htmlname
                    graduado.urlhtmltitulo = urlname
                    graduado.estadonotificacion = 2
                    # fin crear html en la media y guardar url en base
                    graduado.save(request)
                    # envio por correo
                    # correograduado = graduado.inscripcion.persona.emailpersonal()
                    # if IS_DEBUG:
                    #     correograduado = ['marianamadeleineas@gmail.com']
                    # asunto = u"INSIGNIA - " + str(graduado.inscripcion.carrera)
                    # send_html_mail(asunto, "emails/notificar_tituloinsignia_posgrado.html",
                    #                {'sistema': request.session['nombresistema'], 'graduado': graduado,
                    #                 'director': firmacertificado, 'urlhtmlinsignia': urlhtmlinsignia},
                    #                correograduado,
                    #                [], [graduado.rutapdftitulo],
                    #                cuenta=CUENTAS_CORREOS[0][1])
                    if procesomasivo:
                        if not IS_DEBUG:
                            time.sleep(7)
                else:
                    return valida
                lista_correctos.append(f'{graduado.inscripcion.persona.cedula} [{graduado.id}]\n')
        except Exception as ex:
            # transaction.set_rollback(True)
            lista_errores.append(f'{graduado.inscripcion.persona.cedula} [{graduado.id}] error {str(ex)} on line {str(sys.exc_info()[-1].tb_lineno)}\n')
            print(f'{graduado.inscripcion.persona.cedula}[{graduado.id} ] error {str(ex)} on line {str(sys.exc_info()[-1].tb_lineno)}\n')
    print('Titulo/insignia correctos: ', lista_correctos)
    print('lista errores', lista_errores)
    # Para masivo enviar notificacion al sga del administrativo
    if procesomasivo:
        if len(lista_errores) == 0:
            titulonotificacion = f"Proceso exitoso de generación de títulos masivo"
            cuerponotificacion = u"Se generó correctamente el proceso. \nTotal correctos: %s, \nTotal graduados: %s. Generados correctamente: %s" % (
                str(len(lista_correctos)), str(len(graduados)), str(lista_correctos))
        else:
            titulonotificacion = f"Error en el proceso de generación de títulos de graduados de posgrado"
            cuerponotificacion = u"No se generaron %s título(s): \n%s, \nTotal correctos: %s, \nTotal graduados: %s. Generados correctamente: %s" % (
                str(len(lista_errores)), lista_errores, str(len(lista_correctos)), str(len(graduados)), str(lista_correctos))
        # Notifica el resultado del proceso como notificacion en el sga
        notificacion = Notificacion(
            titulo=titulonotificacion,
            cuerpo=cuerponotificacion,
            destinatario=request.session['persona'],
            url=f"/graduados",
            content_type=None,
            object_id=None,
            prioridad=1,
            app_label=request.session['tiposistema'],
            fecha_hora_visible=datetime.now() + timedelta(days=3))
        notificacion.save(request)
    return lista_errores


def solonotificarcorreoinsigniaposgrado(request, graduados, procesomasivo=False, IS_DEBUG=False):
    dominio_sistema = 'http://127.0.0.1:8000'
    if not IS_DEBUG:
        dominio_sistema = 'https://sga.unemi.edu.ec'
    version = datetime.now().strftime('%Y%m%d_%H%M%S%f')
    # Firmante del correo
    firmacertificado = None
    if PersonaDepartamentoFirmas.objects.filter(status=True, departamento=158).exists():
        firmacertificado = PersonaDepartamentoFirmas.objects.filter(status=True, departamento=158).order_by('-id').first()
    lista_correctos = []
    lista_errores = []
    # pruebas local
    if IS_DEBUG and len(graduados) > 2:
        graduados = graduados[:2]

    for graduado in graduados:
        try:
            urlhtmlinsignia = dominio_sistema + graduado.urlhtmltitulo
            # envio por correo
            correograduado = graduado.inscripcion.persona.lista_emails()
            if IS_DEBUG:
                # persona = Persona.objects.get(id=21966)
                # correograduado = persona.lista_emails()
                correograduado = ['pruebasdesarrollo2023@gmail.com']
            asunto = u"INSIGNIA - " + str(graduado.inscripcion.carrera)
            send_html_mail(asunto, "emails/notificar_tituloinsignia_posgrado.html",
                           {'sistema': request.session['nombresistema'], 'graduado': graduado,
                            'director': firmacertificado, 'urlhtmlinsignia': f'{urlhtmlinsignia}?v={version}'},
                           correograduado,
                           [], None,
                           cuenta=CUENTAS_CORREOS[0][1])
            lista_correctos.append(f'{graduado.inscripcion.persona.cedula} [{graduado.id}]')
            # se cambia a estado NOTIFICADO AL CORREO
            graduado.estadonotificacion = 3
            graduado.save(request)
            # if procesomasivo:
            #     if not IS_DEBUG:
            #         time.sleep(5)
        except Exception as ex:
            lista_errores.append(f'Cédula: {graduado.inscripcion.persona.cedula} [{graduado.id}] error: {ex}\n')
            print(f'Cédula: {graduado.inscripcion.persona.cedula} - Código graduado: {graduado.id} - Error: {ex}\n')
    if procesomasivo:
        # notificacion al administrativo en sga
        if len(lista_errores) == 0:
            titulonotificacion = f"Emails enviados exitosamente con los Títulos de posgrado (masivo)"
            cuerponotificacion = u"Se envió los emails correctamente. \nTotal correctos: %s, \nTotal graduados: %s. Emails enviados correctamente: %s" % (
                str(len(lista_correctos)), str(len(graduados)), str(lista_correctos))
        else:
            titulonotificacion = f"Error en el envío de emails con los títulos de posgrado (masivo)"
            cuerponotificacion = u"No se enviaron %s emails: \n%s, \nTotal correctos: %s, \nTotal graduados: %s" % (
                str(len(lista_errores)), lista_errores, str(len(lista_correctos)), str(len(graduados)))
        # Notifica el resultado del proceso como notificacion en el sga
        notificacion = Notificacion(
            titulo=titulonotificacion,
            cuerpo=cuerponotificacion,
            destinatario=request.session['persona'],
            url=f"/graduados",
            content_type=None,
            object_id=None,
            prioridad=1,
            app_label=request.session['tiposistema'],
            fecha_hora_visible=datetime.now() + timedelta(days=3))
        notificacion.save(request)
        # fin notificacion al administrativo sga
    return lista_errores


# convertir de pdf a jpg
def convertir_certificadopdf_a_jpg(qrname, SITE_STORAGE, dominio_sistema, data):
    # ruta jpg
    jpgname = f'{qrname}'
    rutajpg = f'{SITE_STORAGE}/media/qrcode/titulos/{jpgname}.jpg'
    if os.path.isfile(rutajpg):

        os.remove(f'{rutajpg}')
    with open(f'{SITE_STORAGE}/media/qrcode/titulos/{qrname}.pdf', mode='rb') as pdf:
        images = convert_from_bytes(pdf.read(),
                                    output_folder=f'{SITE_STORAGE}/media/qrcode/titulos/',
                                    # first_page = True,
                                    poppler_path=SITE_POPPLER,
                                    fmt="jpg",
                                    single_file=True,
                                    thread_count=1,
                                    # size=(507, 335), # tamaño optimizado
                                    size=(711, 519), # tamaño inicial
                                    output_file=f'{jpgname}'
                                    )
    data['url_jpg'] = dominio_sistema + f'/media/qrcode/titulos/{jpgname}.jpg'
    # data['ruta_jpg'] = rutajpg


# # convertir de pdf a jpg
# def convertir_certificadopdf_a_jpg(qrname, SITE_STORAGE, dominio_sistema, data):
#     # ruta jpg
#     jpgname = f'{qrname}'
#     rutajpg = f'{SITE_STORAGE}/media/qrcode/certificados/{jpgname}.jpg'
#     if os.path.isfile(rutajpg):
#         os.remove(f'{rutajpg}')
#     with open(f'{SITE_STORAGE}/media/qrcode/certificados/{qrname}.pdf', mode='rb') as pdf:
#         images = convert_from_bytes(pdf.read(),
#                                     output_folder=f'{SITE_STORAGE}/media/qrcode/certificados/',
#                                     # first_page = True,
#                                     poppler_path=SITE_POPPLER,
#                                     fmt="jpg",
#                                     single_file=True,
#                                     thread_count=1,
#                                     output_file=f'{jpgname}')
#     data['url_jpg'] = dominio_sistema + f'/media/qrcode/certificados/{jpgname}.jpg'


def actualizarpromediofinal(graduado):
    if ExamenComlexivoGraduados.objects.filter(graduado=graduado, status=True).exists():
        promediotitulacion = ExamenComlexivoGraduados.objects.filter(graduado=graduado, status=True).aggregate(promedio=Avg('examen'))['promedio']
        # cursor = connections['sga_select'].cursor()
        # sql = "select null_to_decimal((" + str(promediotitulacion) + "),2)"
        # cursor.execute(sql)
        # results = cursor.fetchall()
        # for r in results:
        #     campo1 = r[0]
        graduado.promediotitulacion = null_to_decimal(promediotitulacion, 2)
        # graduado.promediotitulacion = campo1
        graduado.save()
