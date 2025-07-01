# -*- coding: UTF-8 -*-
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.http import HttpResponseRedirect, JsonResponse, HttpResponse
from django.shortcuts import render
from django.contrib import messages
from bd.forms import OpcionSistemaForm, ProcesoOpcionSistemaForm, TipoOpcionSistemaForm
from django.template.loader import get_template
from bd.models import InventarioOpcionSistema, ProcesoOpcionSistema, TipoOpcionSistema
from decorators import secure_module
from sga.models import Modulo, PerfilUsuario
from sga.commonviews import adduserdata
from django.db import connection, transaction
import openpyxl
import xlwt
from xlwt import *
import random
import sys
from sga.funciones import MiPaginador, log, generar_nombre, remover_caracteres_especiales_unicode, variable_valor, \
    puede_realizar_accion_afirmativo
from settings import EMPLEADORES_GRUPO_ID
from sga.templatetags.sga_extras import encrypt


@login_required(redirect_field_name='ret', login_url='/loginsga')
# @secure_module
@transaction.atomic()
def view(request):
    data = {}
    adduserdata(request, data)
    persona = request.session['persona']
    if request.method == 'POST':
        action = request.POST['action']

        if action == 'add':
            try:
                with transaction.atomic():
                    form = OpcionSistemaForm(request.POST)
                    if form.is_valid():
                        if not InventarioOpcionSistema.objects.values('id').filter(
                                url=form.cleaned_data['url']).exists():
                            instance = InventarioOpcionSistema(modulo=form.cleaned_data['modulo'],
                                                               nombre=form.cleaned_data['nombre'],
                                                               url=form.cleaned_data['url'],
                                                               descripcion=form.cleaned_data['descripcion'],
                                                               proceso=form.cleaned_data['proceso'],
                                                               tipo=form.cleaned_data['tipo'],
                                                               preguntauxplora=form.cleaned_data['preguntauxplora']
                                                               )
                            if form.cleaned_data['activo']:
                                instance.activo = form.cleaned_data['activo']
                            if form.cleaned_data['posicion_x']:
                                instance.posicion_x = form.cleaned_data['posicion_x']
                            if form.cleaned_data['posicion_y']:
                                instance.posicion_y = form.cleaned_data['posicion_y']
                            if form.cleaned_data['ratio']:
                                instance.ratio = form.cleaned_data['ratio']
                            instance.save(request)
                            if 'archivo' in request.FILES:
                                newfile = request.FILES['archivo']
                                newfile._name = generar_nombre(
                                    'archivo_{}'.format(remover_caracteres_especiales_unicode(instance.nombre)),
                                    newfile._name)
                                instance.archivo = newfile
                                instance.save(request)
                            log(u'Adicionó registro de opción del sistema: %s' % instance, request, "add")
                            messages.success(request, 'Registro guardado con éxito.')
                            return JsonResponse({"result": False}, safe=False)
                        else:
                            raise NameError(u'El registro ya existe. La URL debe ser única.')
                    else:
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                             "mensaje": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": '%s' % ex}, safe=False)

        elif action == 'edit':
            try:
                with transaction.atomic():
                    filtro = InventarioOpcionSistema.objects.get(pk=int(encrypt(request.POST['id'])))
                    f = OpcionSistemaForm(request.POST)
                    if f.is_valid():
                        if not InventarioOpcionSistema.objects.values('id').filter(status=True, url=f.cleaned_data['url']).exclude(
                                pk=filtro.id).exists():
                            filtro.modulo = f.cleaned_data['modulo']
                            filtro.nombre = f.cleaned_data['nombre']
                            filtro.url = f.cleaned_data['url']
                            filtro.descripcion = f.cleaned_data['descripcion']
                            filtro.proceso = f.cleaned_data['proceso']
                            filtro.tipo = f.cleaned_data['tipo']
                            filtro.preguntauxplora = f.cleaned_data['preguntauxplora']
                            if f.cleaned_data['activo']:
                                filtro.activo = f.cleaned_data['activo']
                            if f.cleaned_data['posicion_x']:
                                filtro.posicion_x = f.cleaned_data['posicion_x']
                            if f.cleaned_data['posicion_y']:
                                filtro.posicion_y = f.cleaned_data['posicion_y']
                            if f.cleaned_data['ratio']:
                                filtro.ratio = f.cleaned_data['ratio']
                            filtro.save(request)
                            if 'archivo' in request.FILES:
                                newfile = request.FILES['archivo']
                                newfile._name = generar_nombre(
                                    'archivo_{}'.format(remover_caracteres_especiales_unicode(filtro.nombre)),
                                    newfile._name)
                                filtro.archivo = newfile
                                filtro.save(request)
                            log(u'Editó registro de opción del sistema: %s' % filtro, request, "edit")
                            messages.success(request, 'Registro editado con éxito.')
                            return JsonResponse({"result": False}, safe=False)
                        else:
                            raise NameError('El registro ya existe. La URL debe ser única.')
                    else:
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                             "message": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": '%s' % ex}, safe=False)

        elif action == 'delregistro':
            try:
                with transaction.atomic():
                    instancia = InventarioOpcionSistema.objects.get(pk=int(encrypt(request.POST['id'])))
                    instancia.status = False
                    instancia.save(request)
                    log(u'Eliminó registro de opción del sistema: %s' % instancia, request, "delete")
                    messages.success(request, 'Registro eliminado con éxito.')
                    res_json = {"error": False}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        elif action == 'importarregistros':
            try:
                with transaction.atomic():
                    if not 'archivo_excel' in request.FILES:
                        raise NameError('Carge un archivo excel para ejecutar la acción.')
                    excel = request.FILES['archivo_excel']
                    wb = openpyxl.load_workbook(excel)
                    worksheet = wb.worksheets[0]
                    linea = 1
                    for row in worksheet.iter_rows():
                        currentValues = [str(cell.value) for cell in row]
                        if linea >= 2:
                            if currentValues[0].strip() in ('None', ''):
                                raise NameError(
                                    'REVISAR LINEA [{}],  FORMATO FILAS SIN CÓDIGO DEL MÓDULO, VERIFIQUE QUE EL ARCHIVO NO TENGA COLUMNAS VACIAS AL FINAL.'.format(
                                        linea))
                            if currentValues[1].strip() in ('None', ''):
                                raise NameError(
                                    'REVISAR LINEA [{}],  FORMATO FILAS SIN NOMBRE DE OPCIÓN, VERIFIQUE QUE EL ARCHIVO NO TENGA COLUMNAS VACIAS AL FINAL.'.format(
                                        linea))
                            if currentValues[2].strip() in ('None', ''):
                                raise NameError(
                                    'REVISAR LINEA [{}],  FORMATO FILAS SIN URL, VERIFIQUE QUE EL ARCHIVO NO TENGA COLUMNAS VACIAS AL FINAL.'.format(
                                        linea))
                            if currentValues[3].strip() in ('None', ''):
                                raise NameError(
                                    'REVISAR LINEA [{}],  FORMATO FILAS SIN DESCRIPCIÓN, VERIFIQUE QUE EL ARCHIVO NO TENGA COLUMNAS VACIAS AL FINAL.'.format(
                                        linea))
                            if currentValues[4].strip() in ('None', ''):
                                raise NameError(
                                    'REVISAR LINEA [{}],  FORMATO FILAS SIN CÓDIGO DEL PROCESO, VERIFIQUE QUE EL ARCHIVO NO TENGA COLUMNAS VACIAS AL FINAL.'.format(
                                        linea))
                            if currentValues[5].strip() in ('None', ''):
                                raise NameError(
                                    'REVISAR LINEA [{}],  FORMATO FILAS SIN CÓDIGO DEL TIPO, VERIFIQUE QUE EL ARCHIVO NO TENGA COLUMNAS VACIAS AL FINAL.'.format(
                                        linea))
                            modulo = currentValues[0].replace('\t', '').rstrip() if currentValues[0] != 'None' else None
                            nombre = currentValues[1].replace('\t', '').rstrip() if currentValues[1] != 'None' else None
                            url = currentValues[2].replace('\t', '').rstrip() if currentValues[2] != 'None' else None
                            descripcion = currentValues[3].replace('\t', '').rstrip() if currentValues[
                                                                                             3] != 'None' else None
                            proceso = currentValues[4].replace('\t', '').rstrip() if currentValues[
                                                                                         4] != 'None' else None
                            tipo = currentValues[5].replace('\t', '').rstrip() if currentValues[5] != 'None' else None
                            eModulo, msn = None, ''
                            if modulo:
                                eModulo = Modulo.objects.get(pk=modulo)
                            if proceso:
                                eProceso = ProcesoOpcionSistema.objects.get(pk=proceso)
                            if tipo:
                                eTipo = TipoOpcionSistema.objects.get(pk=tipo)
                            if not InventarioOpcionSistema.objects.values('id').filter(url=url).exists():
                                instance = InventarioOpcionSistema(modulo=eModulo,
                                                                   nombre=nombre,
                                                                   url=url,
                                                                   proceso=eProceso,
                                                                   tipo=eTipo,
                                                                   descripcion=descripcion)
                                instance.save(request)
                                log(u'Adicionó registro de opción del sistema: %s' % instance, request, "add")
                            else:
                                msn = 'Se omitieron registros que ya se encontraban en el sistema.'
                        linea += 1
                    messages.success(request, 'Datos importados correctamente. %s' % msn)
                    return JsonResponse({"result": False}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                print(ex)
                print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
                return JsonResponse({"result": True, "mensaje": '%s' % ex}, safe=False)

        elif action == 'addproceso':
            try:
                with transaction.atomic():
                    form = ProcesoOpcionSistemaForm(request.POST)
                    if form.is_valid():
                        if not ProcesoOpcionSistema.objects.values('id').filter(
                                descripcion=form.cleaned_data['descripcion']).exists():
                            instance = ProcesoOpcionSistema(descripcion=form.cleaned_data['descripcion'])
                            instance.save(request)
                            log(u'Adicionó registro de proceso de opción del sistema: %s' % instance, request, "add")
                            messages.success(request, 'Registro guardado con éxito.')
                            return JsonResponse({"result": False}, safe=False)
                        else:
                            raise NameError(u'El registro ya existe.')
                    else:
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                             "mensaje": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": '%s' % ex}, safe=False)

        elif action == 'editproceso':
            try:
                with transaction.atomic():
                    filtro = ProcesoOpcionSistema.objects.get(pk=int(encrypt(request.POST['id'])))
                    f = ProcesoOpcionSistemaForm(request.POST)
                    if f.is_valid():
                        if not ProcesoOpcionSistema.objects.values('id').filter(
                                descripcion=f.cleaned_data['descripcion']).exclude(pk=filtro.id).exists():
                            filtro.descripcion = f.cleaned_data['descripcion']
                            filtro.save(request)
                            log(u'Editó registro de proceso de opción del sistema: %s' % filtro, request, "edit")
                            messages.success(request, 'Registro editado con éxito.')
                            return JsonResponse({"result": False}, safe=False)
                        else:
                            raise NameError('El registro ya existe.')
                    else:
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                             "message": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": '%s' % ex}, safe=False)

        elif action == 'delproceso':
            try:
                with transaction.atomic():
                    instancia = ProcesoOpcionSistema.objects.get(pk=int(encrypt(request.POST['id'])))
                    instancia.status = False
                    instancia.save(request)
                    log(u'Eliminó registro de proceso de opción del sistema: %s' % instancia, request, "delete")
                    messages.success(request, 'Registro eliminado con éxito.')
                    res_json = {"error": False}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        elif action == 'addtipoopcion':
            try:
                with transaction.atomic():
                    form = TipoOpcionSistemaForm(request.POST)
                    if form.is_valid():
                        if not TipoOpcionSistema.objects.values('id').filter(
                                descripcion=form.cleaned_data['descripcion']).exists():
                            instance = TipoOpcionSistema(descripcion=form.cleaned_data['descripcion'])
                            instance.save(request)
                            log(u'Adicionó registro de tipo de opción del sistema: %s' % instance, request, "add")
                            messages.success(request, 'Registro guardado con éxito.')
                            return JsonResponse({"result": False}, safe=False)
                        else:
                            raise NameError(u'El registro ya existe.')
                    else:
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                             "mensaje": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": '%s' % ex}, safe=False)

        elif action == 'edittipoopcion':
            try:
                with transaction.atomic():
                    filtro = TipoOpcionSistema.objects.get(pk=int(encrypt(request.POST['id'])))
                    f = TipoOpcionSistemaForm(request.POST)
                    if f.is_valid():
                        if not TipoOpcionSistema.objects.values('id').filter(
                                descripcion=f.cleaned_data['descripcion']).exclude(pk=filtro.id).exists():
                            filtro.descripcion = f.cleaned_data['descripcion']
                            filtro.save(request)
                            log(u'Editó registro de tipo de opción del sistema: %s' % filtro, request, "edit")
                            messages.success(request, 'Registro editado con éxito.')
                            return JsonResponse({"result": False}, safe=False)
                        else:
                            raise NameError('El registro ya existe.')
                    else:
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                             "message": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": '%s' % ex}, safe=False)

        elif action == 'deltipoopcion':
            try:
                with transaction.atomic():
                    instancia = TipoOpcionSistema.objects.get(pk=int(encrypt(request.POST['id'])))
                    instancia.status = False
                    instancia.save(request)
                    log(u'Eliminó registro de tipo de opción del sistema: %s' % instancia, request, "delete")
                    messages.success(request, 'Registro eliminado con éxito.')
                    res_json = {"error": False}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        return JsonResponse({"result": "bad", "mensaje": u"Solicitud Incorrecta."})
    else:
        if 'action' in request.GET:
            action = request.GET['action']

            if action == 'add':
                try:
                    form = OpcionSistemaForm()
                    data['action'] = 'add'
                    data['form'] = form
                    template = get_template("adm_sistemas/options/modal/formoption.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'edit':
                try:
                    data['id'] = request.GET['id']
                    data['action'] = 'edit'
                    data['filtro'] = filtro = InventarioOpcionSistema.objects.get(pk=request.GET['id'])
                    form = OpcionSistemaForm(initial={'modulo': filtro.modulo,
                                                      'nombre': filtro.nombre,
                                                      'url': filtro.url,
                                                      'descripcion': filtro.descripcion,
                                                      'proceso': filtro.proceso,
                                                      'tipo': filtro.tipo,
                                                      'archivo': filtro.archivo,
                                                      'preguntauxplora': filtro.preguntauxplora,
                                                      })
                    data['form'] = form
                    template = get_template("adm_sistemas/options/modal/formoption.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'viewperfil':
                try:
                    if InventarioOpcionSistema.objects.values('id').filter(pk=int(request.GET['id'])).exists():
                        opcion = InventarioOpcionSistema.objects.get(pk=int(request.GET['id']))
                        listado = []
                        contador = 0
                        data['modulo'] = u"%s" % opcion.modulo
                        perfiles = PerfilUsuario.objects.filter(persona__usuario__groups__modulogrupo__status=True,
                                                                persona__usuario__groups__modulogrupo__modulos__id=opcion.modulo.id,
                                                                status=True).order_by('persona__usuario').distinct()

                        for perfil in perfiles:
                            if str(perfil) not in listado:
                                if str(perfil) in ['ESTUDIANTE', 'ASPIRANTE ADMISIÓN Y NIVELACIÓN']:
                                    if opcion.modulo.sagest == True:
                                        contador += 1
                                        listado.append('%s' % str(perfil))
                                elif str(perfil) in ['PROFESOR', 'ADMINISTRATIVO']:
                                    if opcion.modulo.sga == True:
                                        contador += 1
                                        listado.append('%s' % str(perfil))
                                elif str(perfil) in ['EMPLEADOR']:
                                    if opcion.modulo.modulogrupo_set.filter(grupos__id__in=[EMPLEADORES_GRUPO_ID]):
                                        contador += 1
                                        listado.append('%s' % str(perfil))
                                elif str(perfil) in ['INSTRUCTOR']:
                                    INSTRUCTOR_GROUP_ID = variable_valor('INSTRUCTOR_GROUP_ID')
                                    if opcion.modulo.modulogrupo_set.filter(grupos__id__in=[INSTRUCTOR_GROUP_ID]):
                                        contador += 1
                                        listado.append('%s' % str(perfil))
                                elif str(perfil) in ['EXTERNO']:
                                    if opcion.modulo.sga == True:
                                        contador += 1
                                        listado.append('%s' % str(perfil))
                                elif str(perfil) in ['ASPIRANTE MAESTRIA']:
                                    if opcion.modulo.posgrado == True:
                                        contador += 1
                                        listado.append('%s' % str(perfil))
                                elif str(perfil) in ['POSTULACIÓN MAESTRIA']:
                                    if opcion.modulo.postulacionposgrado == True:
                                        contador += 1
                                        listado.append('%s' % str(perfil))
                                elif str(perfil) in ['EMPLEO']:
                                    if opcion.modulo.empleo == True:
                                        contador += 1
                                        listado.append('%s' % str(perfil))
                                elif str(perfil) in ['POSTULATE']:
                                    if opcion.modulo.postulate == True:
                                        contador += 1
                                        listado.append('%s' % str(perfil))
                                else:
                                    contador += 1
                                    listado.append('%s' % str(perfil))
                        data['top'] = contador
                        data['listado'] = listado
                        template = get_template("adm_sistemas/options/modal/viewperfil.html")
                        return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'importarregistros':
                try:
                    data['id'] = request.GET['id']
                    data['action'] = 'importarregistros'
                    template = get_template("adm_sistemas/options/modal/importarRegistros.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            elif action == 'viewproceso':
                try:
                    data['title'] = 'Administración de Proceso'
                    filtros, s, url_vars = Q(status=True), request.GET.get('s', ''), ''
                    if s:
                        filtros = filtros & (Q(descripcion__icontains=s))
                        data['s'] = f"{s}"
                        url_vars += f"&s={s}"
                    listado = ProcesoOpcionSistema.objects.filter(filtros).order_by('id')
                    paging = MiPaginador(listado, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['page'] = page
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['eListado'] = page.object_list
                    data['count'] = listado.values("id").count()
                    data['url_vars'] = url_vars
                    return render(request, "adm_sistemas/options/procesoview.html", data)
                except Exception as ex:
                    pass

            elif action == 'addproceso':
                try:
                    form = ProcesoOpcionSistemaForm()
                    data['action'] = 'addproceso'
                    data['form'] = form
                    template = get_template("adm_sistemas/options/modal/forms.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'editproceso':
                try:
                    data['id'] = request.GET['id']
                    data['action'] = 'editproceso'
                    data['filtro'] = filtro = ProcesoOpcionSistema.objects.get(pk=request.GET['id'])
                    form = ProcesoOpcionSistemaForm(initial={'descripcion': filtro.descripcion})
                    data['form'] = form
                    template = get_template("adm_sistemas/options/modal/forms.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'viewtipoopcion':
                try:
                    data['title'] = 'Administración de Tipos de Opción'
                    filtros, s, url_vars = Q(status=True), request.GET.get('s', ''), ''
                    if s:
                        filtros = filtros & (Q(descripcion__icontains=s))
                        data['s'] = f"{s}"
                        url_vars += f"&s={s}"
                    listado = TipoOpcionSistema.objects.filter(filtros).order_by('id')
                    paging = MiPaginador(listado, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['page'] = page
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['eListado'] = page.object_list
                    data['count'] = listado.values("id").count()
                    data['url_vars'] = url_vars
                    return render(request, "adm_sistemas/options/tipoopcionview.html", data)
                except Exception as ex:
                    pass

            if action == 'addtipoopcion':
                try:
                    form = TipoOpcionSistemaForm()
                    data['action'] = 'addtipoopcion'
                    data['form'] = form
                    template = get_template("adm_sistemas/options/modal/forms.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'edittipoopcion':
                try:
                    data['id'] = request.GET['id']
                    data['action'] = 'edittipoopcion'
                    data['filtro'] = filtro = TipoOpcionSistema.objects.get(pk=request.GET['id'])
                    form = TipoOpcionSistemaForm(initial={'descripcion': filtro.descripcion})
                    data['form'] = form
                    template = get_template("adm_sistemas/options/modal/forms.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'excelgeneral':
                try:
                    __author__ = 'Unemi'
                    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on',
                                      num_format_str='#,##0.00')
                    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                    style1 = easyxf(num_format_str='D-MMM-YY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('opciones')
                    response = HttpResponse(content_type="application/ms-excel")
                    response[
                        'Content-Disposition'] = 'attachment; filename=Lista_opciones_sistema' + random.randint(1,
                                                                                                                10000).__str__() + '.xls'
                    columns = [
                        (u"CODIGO", 2000),
                        (u"CODIGO_MODULO", 5000),
                        (u"NOMBRE MODULO", 10000),
                        (u"NOMBRE OPCION", 10000),
                        (u"URL", 16000),
                        (u"DESCRIPCION", 18000),
                        (u"CODIGO_PROCESO", 5000),
                        (u"CODIGO_TIPO", 5000)
                    ]
                    row_num = 0
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    date_format = xlwt.XFStyle()
                    date_format.num_format_str = 'yyyy/mm/dd'
                    listado = InventarioOpcionSistema.objects.filter(status=True).order_by('nombre')
                    row_num = 1
                    i = 0
                    for index, dato in enumerate(listado):
                        campo0 = dato.id
                        campo1 = dato.modulo.id if dato.modulo else ''
                        campo2 = dato.modulo.nombre if dato.modulo else ''
                        campo3 = dato.nombre
                        campo4 = dato.url
                        campo5 = dato.descripcion
                        campo6 = dato.proceso.id if dato.proceso else ''
                        campo7 = dato.tipo.id if dato.tipo else ''

                        ws.write(row_num, 0, campo0, font_style2)
                        ws.write(row_num, 1, campo1, font_style2)
                        ws.write(row_num, 2, campo2, date_format)
                        ws.write(row_num, 3, campo3, font_style2)
                        ws.write(row_num, 4, campo4, font_style2)
                        ws.write(row_num, 5, campo5, font_style2)
                        ws.write(row_num, 6, campo6, font_style2)
                        ws.write(row_num, 7, campo7, font_style2)
                        row_num += 1

                    wb.save(response)
                    return response
                except Exception as ex:
                    pass

            elif action == 'excelgeneralproceso':
                try:
                    __author__ = 'Unemi'
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('tiposopcion')
                    response = HttpResponse(content_type="application/ms-excel")
                    response[
                        'Content-Disposition'] = 'attachment; filename=Lista_procesos_opcion' + random.randint(1,
                                                                                                               10000).__str__() + '.xls'
                    columns = [
                        (u"CODIGO", 3000),
                        (u"DESCRIPCION", 10000)
                    ]
                    row_num = 0
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    date_format = xlwt.XFStyle()
                    date_format.num_format_str = 'yyyy/mm/dd'
                    listado = ProcesoOpcionSistema.objects.filter(status=True).order_by('descripcion')
                    row_num = 1
                    i = 0
                    for index, dato in enumerate(listado):
                        campo0 = dato.id
                        campo1 = dato.descripcion
                        ws.write(row_num, 0, campo0, font_style2)
                        ws.write(row_num, 1, campo1, font_style2)
                        row_num += 1
                    wb.save(response)
                    return response
                except Exception as ex:
                    pass

            return HttpResponseRedirect(request.path)
        else:
            try:
                data['title'] = 'Opciones del Sistema'
                filtros, s, m, t, url_vars, id = Q(status=True), request.GET.get('s', ''), request.GET.get('m',
                                                                                                           '0'), request.GET.get(
                    't', '0'), '', request.GET.get('id', '0')
                eOpciones = InventarioOpcionSistema.objects.values("id").filter(filtros).count()
                if int(id):
                    filtros = filtros & (Q(id=id))
                    data['id'] = f"{id}"
                    url_vars += f"&id={id}"
                if s:
                    filtros = filtros & (Q(nombre__icontains=s) | Q(modulo__nombre__icontains=s) | Q(
                        descripcion__icontains=s) | Q(url__icontains=s) | Q(preguntauxplora__icontains=s))
                    data['s'] = f"{s}"
                    url_vars += f"&s={s}"

                if int(m):
                    filtros = filtros & (Q(modulo_id=m))
                    data['m'] = f"{m}"
                    url_vars += f"&m={m}"

                if int(t):
                    filtros = filtros & (Q(tipo_id=t))
                    data['t'] = f"{t}"
                    url_vars += f"&t={t}"

                eOpciones = InventarioOpcionSistema.objects.filter(filtros).order_by('nombre')
                paging = MiPaginador(eOpciones, 25)
                p = 1
                try:
                    paginasesion = 1
                    if 'paginador' in request.session:
                        paginasesion = int(request.session['paginador'])
                    if 'page' in request.GET:
                        p = int(request.GET['page'])
                    else:
                        p = paginasesion
                    try:
                        page = paging.page(p)
                    except:
                        p = 1
                    page = paging.page(p)
                except:
                    page = paging.page(p)
                request.session['paginador'] = p
                data['paging'] = paging
                data['page'] = page
                data['rangospaging'] = paging.rangos_paginado(p)
                data['eOpciones'] = page.object_list
                data['url_vars'] = url_vars
                data['eModulos'] = Modulo.objects.filter(
                    pk__in=(InventarioOpcionSistema.objects.values_list('modulo_id', flat=True).filter(status=True)),
                    status=True).distinct()
                data['eTipos'] = TipoOpcionSistema.objects.filter(status=True)
                data['tienepermiso'] = puede_realizar_accion_afirmativo(request, 'sga.puede_descargar_db_backup')
                return render(request, "adm_sistemas/options/view.html", data)
            except Exception as ex:
                HttpResponseRedirect(f"/adm_sistemas?info={ex.__str__()}")
