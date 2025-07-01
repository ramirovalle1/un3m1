# -*- coding: UTF-8 -*-
import io
import json
import os
import sys
from datetime import datetime, date, timedelta

from django.core.checks import messages
from openpyxl import workbook as openxl
from openpyxl.styles import Font as openxlFont
from openpyxl.styles.alignment import Alignment as alin
from openpyxl.styles import Border, Side

import openpyxl
import xlwt
from xlwt import *
import random
import pyqrcode, qrcode
from django.db.models import Sum, IntegerField, Min
from django.db.models.functions import Cast
from django.db.models import Count
from googletrans import Translator
from django.contrib.auth.decorators import login_required
from django.db import transaction, connection
from django.db.models import Q, Max
from django.forms import model_to_dict
from django.http import HttpResponseRedirect, JsonResponse, HttpResponse
from django.shortcuts import render
from django.template import Context
from django.template.loader import get_template
from openpyxl import load_workbook

from core.firmar_documentos import firmararchivogenerado
from decorators import secure_module
from sagest.commonviews import secuencia_activos
from sga.tasks import send_html_mail, conectar_cuenta
from posgrado.forms import AsesorComercialForm
from sagest.forms import ActivoFijoForm, InformeBajaForm, DetalleInformeBajaForm, \
    ImportarArchivoXLSForm, TraspasoActivoForm, BajaActivoForm, DetalleBajaActivoForm, AsignacionActivoForm, \
    TrasladoMantenimientoForm, MantenimientosActivosGarantiaForm, \
    DetalleNoIdentificadoForm, ActivosFijosForm, \
    DetalleTarjetaControlForm, ActasForm, ConsultacatalogoForm, ConsultausuarioForm, ConstatacionForm, \
    TraspasoActivoCustodioForm, ExportacionForm, EstadoBienForm, CondicionBienForm, EdificioForm, ArchivoActivoBajaForm, \
    ReporteEdificioFrom, TareaMantenimientoFrom, TipoBajaForm, TareaMantenimientoDaniosForm, \
    MantenimientosActivosPreventivosForm, GrupoBienForm, DirectorResponsableBajaForm, DirectorResponsableBaja2Form, \
    InformeBajaForm2, ResponsableActivoForm, InventarioActivoForm, PeriodoInventarioATForm, EvidenciaPeriodoInventarioTecnologicoForm, \
    EstadosGeneralesInventarioATForm,CronogramaPersonaInventarioForm, ConstatacionFisicaATForm
from settings import RESPONSABLE_BIENES_ID, MEDIA_ROOT, ASISTENTE_BODEGA_ID, SITE_ROOT, SITE_STORAGE, MEDIA_URL, DEBUG
from sagest.models import ActivoFijo, ArchivoActivoFijo, HdDetalle_Incidente, MantenimientosActivosGarantia, \
    DetalleMantenimientosActivosGarantia, PersonaDepartamentoFirmas, \
    TraspasoActivo, DetalleTraspasoActivo, BajaActivo, DetalleBajaActivo, Ubicacion, \
    TrasladoMantenimiento, DetalleTrasladoMantenimiento, ConstatacionFisica, TIPO_MANTENIMIENTO, \
    DetalleConstatacionFisica, DetalleNoIdentificado, CuentaContable, DetalleMantenimiento, TareasActivosPreventivos, \
    CatalogoBien, ClaseVehiculo, TipoVehiculo, Color, EstadoProducto, OrigenIngreso, TipoDocumento, TarjetaControl, \
    ExportacionesActivos, Edificio, EstadoBien, CondicionBien, RangoVidaUtil, GruposCategoria, \
    MantenimientosActivosPreventivos, MantenimientoGruCategoria, TipoBaja, InformeActivoBaja, DetalleInformeActivoBaja, \
    MantenimientoGruCategoriaGarantiaLimp, MantenimientoGruCategoriaGarantiaErr, TareasActivosPreventivosGarantiaLimp, \
    TareasActivosPreventivosGarantiaErr, MantenimientoGruDanios, TareasActivosPreventivosDanios, \
    PiezaParteActivosPreventivos, HdPiezaPartes, ESTADO_DANIO, DirectorResponsableBaja, HistorialTraspaso, \
    SolicitudActivos, \
    ActivoFijoInventarioTecnologico, ActivoTecnologico, PeriodoInventarioAT, AsignacionCierreInventarioAT, \
    EvidenciaPeriodoInventarioTecnologico, \
    EstadosGeneralesInventarioAT, InventarioATEstadosGenerales, CronogramaPersonaConstatacionAT, \
    DetalleConstatacionFisicaActivoTecnologico, ComponenteCatalogoActivo, DetalleCatalogoComponenteConstatacionAT, \
    ESTADO_CONSTATACION_AT, HdBloqueUbicacion, HistorialDocumentosFirmadosConstatacionAT
from sga.commonviews import adduserdata, obtener_reporte
from sga.forms import ReservasCraiSolicitarAutoridadForm
from sga.templatetags.sga_extras import encrypt
from sga.funciones import MiPaginador, generar_nombre, log, convertir_fecha, \
    puede_realizar_accion, remover_caracteres_especiales_unicode, null_to_decimal, notificacion, bad_json
from sga.models import Persona, CUENTAS_CORREOS, VisitasBiblioteca, Reporte, Notificacion
from sga.reportes import elimina_tildes, run_report_v1
from sga.funcionesxhtml2pdf import conviert_html_to_pdf, conviert_html_to_pdfsave, conviert_html_to_pdfsaveinformeinventarioactivostecnologicos, \
    conviert_html_to_pdfsaveqrcertificado, conviert_html_to_pdf_name, \
    conviert_html_to_pdfsaveinformeactivo
from sga.excelbackground import reporte_activos_tecnologicos_constatados_openxl_background


unicode = str


@login_required(redirect_field_name='ret', login_url='/loginsagest')
@secure_module
@transaction.atomic()
def view(request):
    data = {}
    adduserdata(request, data)
    usuario = request.user
    hoy=datetime.now()
    data['personasesion'] = personaactivo = request.session['persona']
    h = 'http' if DEBUG else 'https'
    base_url = request.META['HTTP_HOST']
    data['DOMINIO_DEL_SISTEMA'] = dominio_sistema = f"{h}://{unicode(base_url)}"
    if request.method == 'POST':
        action = request.POST['action']

        if action == 'add':
            try:
                f = ActivoFijoForm(request.POST)
                if f.is_valid():
                    if f.cleaned_data['responsable'] or f.cleaned_data['custodio'] or f.cleaned_data['ubicacion']:
                        if not f.cleaned_data['responsable'] or not f.cleaned_data['custodio'] or not f.cleaned_data[
                            'ubicacion']:
                            return JsonResponse(
                                {"result": "bad", "mensaje": u"Debe seleccionar Responsable, Custodio y Ubicación."})
                    activo = ActivoFijo(codigogobierno=f.cleaned_data['codigogobierno'],
                                        codigointerno=f.cleaned_data['codigointerno'],
                                        fechaingreso=f.cleaned_data['fechaingreso'],
                                        observacion=f.cleaned_data['observacion'],
                                        costo=f.cleaned_data['costo'],
                                        serie=f.cleaned_data['serie'],
                                        descripcion=f.cleaned_data['descripcion'],
                                        modelo=f.cleaned_data['modelo'],
                                        marca=f.cleaned_data['marca'],
                                        tipocomprobante=f.cleaned_data['tipocomprobante'],
                                        numerocomprobante=f.cleaned_data['numerocomprobante'],
                                        fechacomprobante=f.cleaned_data['fechacomprobante'],
                                        deprecia=f.cleaned_data['deprecia'],
                                        vidautil=f.cleaned_data['vidautil'],
                                        estructuraactivo=f.cleaned_data['estructuraactivo'],
                                        clasebien=f.cleaned_data['clasebien'],
                                        catalogo_id=f.cleaned_data['catalogo'],
                                        origeningreso=f.cleaned_data['origeningreso'],
                                        tipodocumentorespaldo=f.cleaned_data['tipodocumentorespaldo'],
                                        clasedocumentorespaldo=f.cleaned_data['clasedocumentorespaldo'],
                                        estado=f.cleaned_data['estado'],
                                        cuentacontable=f.cleaned_data['cuentacontable'],
                                        origenregistro=1,
                                        tipoproyecto=f.cleaned_data['tipoproyecto'])
                    activo.save(request)
                    if f.cleaned_data['responsable']:
                        activo.ubicacion = f.cleaned_data["ubicacion"]
                        activo.responsable_id = f.cleaned_data["responsable"]
                        activo.custodio_id = f.cleaned_data["custodio"]
                        traspaso = TraspasoActivo(usuariobienrecibe_id=f.cleaned_data['responsable'],
                                                  custodiobienrecibe_id=f.cleaned_data['custodio'],
                                                  ubicacionbienrecibe=f.cleaned_data['ubicacion'],
                                                  responsablebienes_id=RESPONSABLE_BIENES_ID,
                                                  fecha=datetime.now(),
                                                  tipo=1)
                        traspaso.save(request)
                        log(u'Adicionó traspaso: %s' % traspaso, request, "add")
                        secuencia = secuencia_activos(request)
                        secuencia.numeroasignacion += 1
                        secuencia.save(request)
                        traspaso.numero = secuencia.numeroasignacion
                        traspaso.save(request)
                        detalle = DetalleTraspasoActivo(codigotraspaso=traspaso,
                                                        activo=activo)
                        detalle.save(request)
                    if activo.catalogo.tipobien.id in [1, 3, 5]:
                        activo.color = f.cleaned_data['color']
                        activo.material = f.cleaned_data['material']
                        activo.dimensiones = f.cleaned_data['dimensiones']
                    elif activo.catalogo.tipobien.id == 2:
                        activo.colorprimario = f.cleaned_data['colorprimario']
                        activo.colorsecundario = f.cleaned_data['colorsecundario']
                        activo.clasevehiculo = f.cleaned_data['clasevehiculo']
                        activo.tipovehiculo = f.cleaned_data['tipovehiculo']
                        activo.numerochasis = f.cleaned_data['numerochasis']
                        activo.numeromotor = f.cleaned_data['numeromotor']
                        activo.placa = f.cleaned_data['placa']
                        activo.aniofabricacion = f.cleaned_data['aniofabricacion']
                    elif activo.catalogo.tipobien.id == 4:
                        activo.propietario = f.cleaned_data['propietario']
                        activo.codigocatastral = f.cleaned_data['codigocatastral']
                        activo.numeropredio = f.cleaned_data['numeropredio']
                        activo.valoravaluo = f.cleaned_data['valoravaluo']
                        activo.anioavaluo = f.cleaned_data['anioavaluo']
                        activo.areapredio = f.cleaned_data['areapredio']
                        activo.areaconstruccion = f.cleaned_data['areaconstruccion']
                        activo.pisos = f.cleaned_data['pisos']
                        activo.provincia = f.cleaned_data['provincia']
                        activo.canton = f.cleaned_data['canton']
                        activo.parroquia = f.cleaned_data['parroquia']
                        activo.zona = f.cleaned_data['zona']
                        activo.nomenclatura = f.cleaned_data['nomenclatura']
                        activo.sector = f.cleaned_data['sector']
                        activo.direccion = f.cleaned_data['direccion']
                        activo.escritura = f.cleaned_data['escritura']
                        activo.fechaescritura = f.cleaned_data['fechaescritura']
                        activo.notaria = f.cleaned_data['notaria']
                        activo.beneficiariocontrato = f.cleaned_data['beneficiariocontrato']
                        activo.fechacontrato = f.cleaned_data['fechacontrato']
                        activo.duracioncontrato = f.cleaned_data['duracioncontrato']
                        activo.montocontrato = f.cleaned_data['montocontrato']
                    else:
                        activo.titulo = f.cleaned_data['titulo']
                        activo.autor = f.cleaned_data['autor']
                        activo.editorial = f.cleaned_data['editorial']
                        activo.fechaedicion = f.cleaned_data['fechaedicion']
                        activo.numeroedicion = f.cleaned_data['numeroedicion']
                        activo.clasificacionbibliografica = f.cleaned_data['clasificacionbibliografica']
                    activo.save(request)
                    log(u'Adicionó Activo: %s' % activo, request, "add")
                    cuenta = CuentaContable.objects.get(pk=int(activo.cuentacontable_id))
                    cuenta.activosfijos = True
                    cuenta.save(request)
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'addingresoinformebaja':
            try:
                tipo = int(request.POST['tipo'])
                if InformeActivoBaja.objects.filter(tipoinforme=tipo,
                                                    activofijo_id=int(encrypt(request.POST['id']))).exists():
                    return JsonResponse({"result": "bad", "mensaje": u"El informe ya existe."})

                if tipo == 1:
                    f = InformeBajaForm(request.POST)

                else:
                    f = InformeBajaForm2(request.POST)

                if f.is_valid():
                    if not f.cleaned_data['solicita']:
                        return JsonResponse({"result": "bad", "mensaje": u"Debe seleccionar solicitante."})
                    informebaja = InformeActivoBaja(activofijo_id=int(encrypt(request.POST['id'])),
                                                    solicita_id=f.cleaned_data['solicita'],
                                                    responsable_id=f.cleaned_data['responsable'],
                                                    conclusion=f.cleaned_data['conclusion'],
                                                    estado=f.cleaned_data['estado'],
                                                    estadouso=f.cleaned_data['estadouso'],
                                                    bloque=f.cleaned_data['bloque'],
                                                    detallerevision=f.cleaned_data['detallerevision'],
                                                    tipoinforme=tipo)
                    informebaja.save(request)

                    if tipo == 2:
                        informebaja.departamento = f.cleaned_data['departamento']
                        informebaja.gestion = f.cleaned_data['gestion']
                        informebaja.save(request)

                    if 'lista_items4' in request.POST:
                        listadetalle = json.loads(request.POST['lista_items4'])
                        if listadetalle:
                            for lisdet in listadetalle:
                                ingresodetalle = DetalleInformeActivoBaja(informebaja=informebaja,
                                                                          detalle=lisdet['listadetalle'])
                                ingresodetalle.save(request)

                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'edit':
            try:
                activo = ActivoFijo.objects.get(pk=request.POST['id'])
                f = ActivoFijoForm(request.POST)
                if f.is_valid():
                    activo.observacion = f.cleaned_data['observacion']
                    activo.serie = f.cleaned_data['serie']
                    activo.descripcion = f.cleaned_data['descripcion']
                    activo.modelo = f.cleaned_data['modelo']
                    activo.marca = f.cleaned_data['marca']
                    activo.tipocomprobante = f.cleaned_data['tipocomprobante']
                    activo.numerocomprobante = f.cleaned_data['numerocomprobante']
                    activo.clasebien = f.cleaned_data['clasebien']
                    activo.origeningreso = f.cleaned_data['origeningreso']
                    activo.tipodocumentorespaldo = f.cleaned_data['tipodocumentorespaldo']
                    activo.clasedocumentorespaldo = f.cleaned_data['clasedocumentorespaldo']
                    activo.estado = f.cleaned_data['estado']
                    activo.cuentacontable = f.cleaned_data['cuentacontable']
                    activo.tipoproyecto = f.cleaned_data['tipoproyecto']
                    activo.ubicacion = f.cleaned_data['ubicacion']
                    if activo.custodio == None:
                        activo.custodio = Persona.objects.get(pk=f.cleaned_data['custodio'], status=True)
                    if activo.responsable == None:
                        activo.responsable = Persona.objects.get(pk=f.cleaned_data['responsable'], status=True)
                    if not activo.codigogobierno:
                        activo.codigogobierno = f.cleaned_data['codigogobierno']
                    if not activo.en_uso() and not activo.subidogobierno:
                        activo.fechaingreso = f.cleaned_data['fechaingreso']
                        activo.estructuraactivo = f.cleaned_data['estructuraactivo']
                        activo.costo = f.cleaned_data['costo']
                        activo.fechacomprobante = f.cleaned_data['fechacomprobante']
                    if not activo.en_uso() or request.user.has_perm("sagest.puede_modificar_depreciacion"):
                        activo.deprecia = f.cleaned_data['deprecia']
                        activo.vidautil = f.cleaned_data['vidautil']
                        catalago = activo.catalogo.id
                        if f.cleaned_data['catalogo']:
                            catalogo = CatalogoBien.objects.get(id=int(f.cleaned_data['catalogo']))
                        else:
                            catalogo = CatalogoBien.objects.get(id=activo.catalogo.id)
                        activo.catalogo = catalogo
                    activo.save(request)
                    if activo.catalogo.tipobien.id in [1, 3, 5]:
                        activo.color = f.cleaned_data['color']
                        activo.material = f.cleaned_data['material']
                        activo.dimensiones = f.cleaned_data['dimensiones']
                    elif activo.catalogo.tipobien.id == 2:
                        activo.colorprimario = f.cleaned_data['colorprimario']
                        activo.colorsecundario = f.cleaned_data['colorsecundario']
                        activo.clasevehiculo = f.cleaned_data['clasevehiculo']
                        activo.tipovehiculo = f.cleaned_data['tipovehiculo']
                        activo.numerochasis = f.cleaned_data['numerochasis']
                        activo.numeromotor = f.cleaned_data['numeromotor']
                        activo.placa = f.cleaned_data['placa']
                        activo.aniofabricacion = f.cleaned_data['aniofabricacion']
                    elif activo.catalogo.tipobien.id == 4:
                        activo.propietario = f.cleaned_data['propietario']
                        activo.codigocatastral = f.cleaned_data['codigocatastral']
                        activo.numeropredio = f.cleaned_data['numeropredio']
                        activo.valoravaluo = f.cleaned_data['valoravaluo']
                        activo.anioavaluo = f.cleaned_data['anioavaluo']
                        activo.areapredio = f.cleaned_data['areapredio']
                        activo.areaconstruccion = f.cleaned_data['areaconstruccion']
                        activo.pisos = f.cleaned_data['pisos']
                        activo.provincia = f.cleaned_data['provincia']
                        activo.canton = f.cleaned_data['canton']
                        activo.parroquia = f.cleaned_data['parroquia']
                        activo.zona = f.cleaned_data['zona']
                        activo.nomenclatura = f.cleaned_data['nomenclatura']
                        activo.sector = f.cleaned_data['sector']
                        activo.direccion = f.cleaned_data['direccion']
                        activo.escritura = f.cleaned_data['escritura']
                        activo.fechaescritura = f.cleaned_data['fechaescritura']
                        activo.notaria = f.cleaned_data['notaria']
                        activo.beneficiariocontrato = f.cleaned_data['beneficiariocontrato']
                        activo.fechacontrato = f.cleaned_data['fechacontrato']
                        activo.duracioncontrato = f.cleaned_data['duracioncontrato']
                        activo.montocontrato = f.cleaned_data['montocontrato']
                    else:
                        activo.titulo = f.cleaned_data['titulo']
                        activo.autor = f.cleaned_data['autor']
                        activo.editorial = f.cleaned_data['editorial']
                        activo.fechaedicion = f.cleaned_data['fechaedicion']
                        activo.numeroedicion = f.cleaned_data['numeroedicion']
                        activo.clasificacionbibliografica = f.cleaned_data['clasificacionbibliografica']
                    activo.save(request)
                    log(u'Editó Activo: %s' % activo, request, "edit")
                    fechaultimadeprec = ActivoFijo.objects.aggregate(fecha=Max('fechaultimadeprec'))['fecha']
                    activo.depreciar_activo(fechaultimadeprec, request)
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'editingresoinformebaja':
            try:
                informebaja = InformeActivoBaja.objects.get(pk=int(request.POST['id']))
                f = InformeBajaForm2(request.POST)
                if informebaja.tipoinforme == 1:
                    f = InformeBajaForm(request.POST)
                if f.is_valid():
                    if not f.cleaned_data['solicita']:
                        return JsonResponse({"result": "bad", "mensaje": u"Debe seleccionar solicitante."})
                    informebaja.solicita_id = f.cleaned_data['solicita']
                    informebaja.responsable_id = f.cleaned_data['responsable']
                    informebaja.detallerevision = f.cleaned_data['detallerevision']
                    informebaja.bloque = f.cleaned_data['bloque']
                    informebaja.estadouso = f.cleaned_data['estadouso']
                    informebaja.estado = f.cleaned_data['estado']
                    informebaja.conclusion = f.cleaned_data['conclusion']
                    if informebaja.tipoinforme == 2:
                        informebaja.departamento = f.cleaned_data['departamento']
                        informebaja.gestion = f.cleaned_data['gestion']
                    informebaja.save(request)

                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'additemdetalleinformebaja':
            try:
                detalleinf = DetalleInformeActivoBaja(informebaja_id=request.POST['idactivoinformebaja'],
                                                      detalle=request.POST['descripcion'])
                detalleinf.save(request)
                log(u'Adicinó detalle de informe de baja: %s' % detalleinf, request, "add")
                return JsonResponse({"result": "ok", 'codigoinformedet': detalleinf.id})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'listainformebaja':
            try:
                detalleinf = DetalleInformeActivoBaja.objects.get(pk=request.POST['id'])
                descripcion = detalleinf.detalle
                idobjetivo = detalleinf.id
                return JsonResponse({"result": "ok", 'descripcion': descripcion, 'codigorai': idobjetivo})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'edititemdetalleinformebaja':
            try:
                detalleinf = DetalleInformeActivoBaja.objects.get(pk=request.POST['codigoitemdetalleinformebaja'])
                detalleinf.detalle = request.POST['descripcion']
                detalleinf.save(request)
                log(u'Modifico detalle informe de baja: %s ' % detalleinf, request, "edit")
                return JsonResponse({"result": "ok", "descripcion": detalleinf.detalle})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'itemdetalleinformebaja':
            try:
                detalleinf = DetalleInformeActivoBaja.objects.get(pk=request.POST['id'])
                descripcion = detalleinf.detalle
                idobjetivo = detalleinf.id
                return JsonResponse({"result": "ok", 'descripcion': descripcion, 'codigorai': idobjetivo})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'eliminaritemdetalleinformebaja':
            try:
                itemdetalle = DetalleInformeActivoBaja.objects.get(pk=request.POST['idcodigodet'])
                log(u'Eliminó item detalle informe baja: %s' % itemdetalle, request, "del")
                itemdetalle.delete()
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'editmantenimientogarantia':
            try:
                mangarantia = MantenimientosActivosGarantia.objects.get(pk=int(request.POST['id']))
                limp = request.POST['id_limpiar']
                danio = request.POST['id_danio']
                if 'archivo' in request.FILES:
                    d = request.FILES['archivo']
                    if d.size > 2194304:
                        return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 2 Mb."})
                    else:
                        newfiles = request.FILES['archivo']
                        newfilesd = newfiles._name
                        ext = newfilesd[newfilesd.rfind("."):]
                        if not (ext == '.pdf'):
                            return JsonResponse({"result": "bad", "mensaje": u"Error, solo archivos .pdf"})
                if 'arcusen' in request.FILES:
                    d = request.FILES['arcusen']
                    if d.size > 2194304:
                        return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 2 Mb."})
                    else:
                        newfiles = request.FILES['arcusen']
                        newfilesd = newfiles._name
                        ext = newfilesd[newfilesd.rfind("."):]
                        if not (ext == '.pdf'):
                            return JsonResponse({"result": "bad", "mensaje": u"Error, solo archivos .pdf"})
                f = MantenimientosActivosGarantiaForm(request.POST, request.FILES)
                if f.is_valid():
                    # mangarantia.activofijo_id = f.cleaned_data['activofijo']
                    mangarantia.proveedor_id = f.cleaned_data['proveedor']
                    mangarantia.valor = f.cleaned_data['valor']
                    mangarantia.numreporte = f.cleaned_data['numreporte']
                    mangarantia.fechainicio = f.cleaned_data['fechainicio']
                    mangarantia.horamax = f.cleaned_data['horamax']
                    mangarantia.minutomax = f.cleaned_data['minutomax']
                    mangarantia.estfrec = f.cleaned_data['estfrec']
                    mangarantia.estfent = f.cleaned_data['estfent']
                    mangarantia.observacion = f.cleaned_data['observacion']
                    mangarantia.estusu = f.cleaned_data['estusu']
                    mangarantia.save(request)
                    if 'archivo' in request.FILES:
                        newfile = request.FILES['archivo']
                        newfile._name = generar_nombre("archivo_garantia", newfile._name)
                        mangarantia.archivo = newfile
                        # capacitacion.tiempo = f.cleaned_data['tiempo']
                        mangarantia.save(request)
                    if 'arcusen' in request.FILES:
                        newfile = request.FILES['arcusen']
                        newfile._name = generar_nombre("archivo_entusuario", newfile._name)
                        mangarantia.arcusen = newfile
                        # capacitacion.tiempo = f.cleaned_data['tiempo']
                        mangarantia.save(request)

                    TareasActivosPreventivosGarantiaLimp.objects.filter(mantenimiento=mangarantia).delete()
                    elementoslimp = limp.split(',')
                    for elemento in elementoslimp:
                        detalle = TareasActivosPreventivosGarantiaLimp(mantenimiento=mangarantia, grupos_id=elemento)
                        detalle.save(request)

                    TareasActivosPreventivosGarantiaErr.objects.filter(mantenimiento=mangarantia).delete()
                    elementoserr = danio.split(',')
                    for elemento in elementoserr:
                        detalle = TareasActivosPreventivosGarantiaErr(mantenimiento=mangarantia, grupos_id=elemento)
                        detalle.save(request)
                    log(u'Modifico mantenimiento garantia: %s' % mangarantia, request, "edit")
                    return JsonResponse({'result': 'ok'})
                else:
                    return JsonResponse({"result": "bad", "mensaje": u"Error, al guardar los datos."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': u'Error al guardar los datos'})

        elif action == 'asignar':
            try:
                f = AsignacionActivoForm(request.POST)
                if f.is_valid():
                    items = json.loads(request.POST['lista_items1'])
                    if len(items) == 0:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Debe seleccionar activos para poder asignar"})
                    traspaso = TraspasoActivo(usuariobienrecibe=f.cleaned_data['usuariobienrecibe'],
                                              custodiobienrecibe=f.cleaned_data['custodiobienrecibe'],
                                              ubicacionbienrecibe=f.cleaned_data['ubicacionbienrecibe'],
                                              fecha=datetime.now(),
                                              tipo=1)
                    traspaso.save(request)
                    secuencia = secuencia_activos(request)
                    secuencia.numeroasignacion += 1
                    secuencia.save(request)
                    traspaso.numero = secuencia.numeroasignacion
                    traspaso.save(request)
                    for d in items:
                        activo = ActivoFijo.objects.get(pk=int(d['id']))
                        detalle = DetalleTraspasoActivo(codigotraspaso=traspaso,
                                                        activo=activo)
                        detalle.save(request)
                        activo.actualiza_responsable()

                    log(u'Asigno Activo: %s' % traspaso, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'listatipomantenimiento':
            try:
                listatareas = MantenimientoGruCategoria.objects.filter(grupocategoria_id=request.POST['id_tipo'],
                                                                       status=True, activo=True)
                lista = []
                for lis in listatareas:
                    lista.append([lis.id, lis.descripcion])
                return JsonResponse({'result': 'ok', 'lista': lista})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'listatipomantenimientosgdanio':
            try:
                listatareas = MantenimientoGruDanios.objects.filter(grupocategoria_id=request.POST['id_tipo'],
                                                                    status=True, activo=True)
                lista = []
                for lis in listatareas:
                    lista.append([lis.id, lis.descripcion])
                return JsonResponse({'result': 'ok', 'lista': lista})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'afmarcaymodelo':
            try:
                listatareas = ActivoFijo.objects.filter(id=request.POST['id_tipo'], status=True)
                lista = []
                for lis in listatareas:
                    lista.append([lis.marca, lis.modelo])
                return JsonResponse({'result': 'ok', 'lista': lista})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'addmantenimiento':
            try:
                form = MantenimientosActivosPreventivosForm(request.POST, request.FILES)
                if 'archivo' in request.FILES:
                    newfile = request.FILES['archivo']
                    if newfile:
                        newfilesd = newfile._name
                        ext = newfilesd[newfilesd.rfind("."):]
                        if not ext.lower() == '.pdf':
                            return JsonResponse({"result": "bad", "mensaje": u"Error, solo archivos .pdf."})
                        if newfile.size > 2194304:
                            return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 2 Mb."})
                        if newfile:
                            newfile._name = generar_nombre("archivo_malla_", newfile._name)
                if form.is_valid():
                    id_cpu = request.POST['cpu']
                    id_tipact = request.POST['tipoactivo']
                    estusu = form.cleaned_data['estusu']
                    fecha = form.cleaned_data['fecha']
                    horamax = form.cleaned_data['horamax']
                    minutomax = form.cleaned_data['minutomax']
                    estfrec = form.cleaned_data['estfrec']
                    estfent = form.cleaned_data['estfent']
                    marca = form.cleaned_data['marca']
                    modelo = form.cleaned_data['modelo']
                    bsugiere = form.cleaned_data['bsugiere']
                    dsugiere = form.cleaned_data['dsugiere']
                    observacion = form.cleaned_data['observacion']
                    limp = request.POST['id_limp']
                    # danio = request.POST['id_dani']
                    piezaparte = request.POST.getlist('piezaparte[]')
                    caracteristica = request.POST.getlist('catacteristica[]')
                    danio = request.POST.getlist('danio[]')
                    daniodes = request.POST.getlist('danioop[]')

                    mantenimientosactivos = MantenimientosActivosPreventivos(activofijo_id=id_cpu,
                                                                             tipoactivo_id=id_tipact,
                                                                             estusu=estusu,
                                                                             fecha=fecha,
                                                                             horamax=horamax,
                                                                             minutomax=minutomax,
                                                                             funcionarecibe=estfrec,
                                                                             funcionaentrega=estfent,
                                                                             marca=marca,
                                                                             modelo=modelo,
                                                                             sbequipo=bsugiere,
                                                                             descbaja=dsugiere,
                                                                             observaciones=observacion,
                                                                             nuevo=True)
                    mantenimientosactivos.save(request)
                    limp = limp.split(',')
                    for elemento in limp:
                        detalle = TareasActivosPreventivos(mantenimiento=mantenimientosactivos,
                                                           grupos_id=elemento)
                        detalle.save(request)
                    c = 0
                    while c < len(danio):
                        detalle = TareasActivosPreventivosDanios(mantenimiento=mantenimientosactivos,
                                                                 grupos_id=int(MantenimientoGruDanios.objects.get(
                                                                     descripcion=danio[c], grupocategoria_id=id_tipact,
                                                                     status=True).id),
                                                                 estadodanio=daniodes[c])
                        detalle.save(request)
                        c += 1
                    counter = 0
                    while counter < len(piezaparte):
                        detalle = PiezaParteActivosPreventivos(mantenimiento=mantenimientosactivos,
                                                               piezaparte_id=int(HdPiezaPartes.objects.get(
                                                                   descripcion=piezaparte[counter],
                                                                   grupocategoria_id=id_tipact, status=True).id),
                                                               descripcion=caracteristica[counter])
                        detalle.save(request)
                        counter += 1
                    if 'archivo' in request.FILES:
                        newfile = request.FILES['archivo']
                        if newfile:
                            newfile._name = generar_nombre("evidencia_comunicado", newfile._name)
                            mantenimientosactivos.archivo = newfile
                            mantenimientosactivos.save(request)
                    log(u'Asigno mantenimiento preventivo Activo: %s' % mantenimientosactivos, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'editmantenimiento':
            try:
                f = AsignacionActivoForm(request.POST)
                if f.is_valid():
                    id_activoman = request.POST['id_activoman']
                    id_monitor = request.POST['id_monitor']
                    id_mouse = request.POST['id_mouse']
                    id_teclado = request.POST['id_teclado']
                    id_fechaingreso = request.POST['id_fechaingreso']
                    id_mantenimiento = request.POST['id_mantenimiento']
                    id_tipo = request.POST['id_tipo']
                    id_procesador = request.POST['id_procesador']
                    id_memoria = request.POST['id_memoria']
                    id_discoduro = request.POST['id_discoduro']
                    id_particiones = request.POST['id_particiones']
                    id_sistemaoperativo = request.POST['id_sistemaoperativo']
                    id_arquitectura = request.POST['id_arquitectura']
                    id_service = request.POST['id_service']
                    id_observacion = request.POST['id_observacion']
                    lista = request.POST['lista']
                    id_recibe = request.POST['id_recibe']
                    recibeid = False
                    if id_recibe == '1':
                        recibeid = True
                    id_entrega = request.POST['id_entrega']
                    entregaid = False
                    if id_entrega == '1':
                        entregaid = True
                    mantenimientosactivos = MantenimientosActivosPreventivos.objects.get(pk=id_activoman)
                    mantenimientosactivos.tipomantenimiento = id_mantenimiento
                    monitor = id_monitor
                    mantenimientosactivos.mouse = id_mouse
                    mantenimientosactivos.teclado = id_teclado
                    mantenimientosactivos.procesador = id_procesador
                    mantenimientosactivos.memoria = id_memoria
                    mantenimientosactivos.discoduro = id_discoduro
                    mantenimientosactivos.particiones = id_particiones
                    mantenimientosactivos.sistemaoperativo = id_sistemaoperativo
                    mantenimientosactivos.arquitectura = id_arquitectura
                    mantenimientosactivos.service = id_service
                    mantenimientosactivos.fecha = id_fechaingreso
                    mantenimientosactivos.tipoactivo_id = id_tipo
                    mantenimientosactivos.funcionarecibe = recibeid
                    mantenimientosactivos.funcionaentrega = entregaid
                    mantenimientosactivos.observaciones = id_observacion
                    mantenimientosactivos.save(request)

                    TareasActivosPreventivos.objects.filter(mantenimiento=mantenimientosactivos).delete()
                    elementos = lista.split(',')
                    for elemento in elementos:
                        detalle = TareasActivosPreventivos(mantenimiento=mantenimientosactivos,
                                                           grupos_id=elemento)
                        detalle.save(request)
                    log(u'Editó mantenimiento preventivo Activo: %s' % mantenimientosactivos, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'editmantenimientov2':
            try:
                form = MantenimientosActivosPreventivosForm(request.POST, request.FILES)
                if form.is_valid():
                    eli_arc = request.POST['archivo-clear'] if 'archivo-clear' in request.POST else ''
                    id_activoman = request.POST['id']
                    id_cpu = request.POST['cpu']
                    id_tipact = request.POST['tipoactivo']
                    estusu = form.cleaned_data['estusu']
                    fecha = form.cleaned_data['fecha']
                    horamax = form.cleaned_data['horamax']
                    minutomax = form.cleaned_data['minutomax']
                    estfrec = form.cleaned_data['estfrec']
                    estfent = form.cleaned_data['estfent']
                    marca = form.cleaned_data['marca']
                    modelo = form.cleaned_data['modelo']
                    bsugiere = form.cleaned_data['bsugiere']
                    dsugiere = form.cleaned_data['dsugiere']
                    observacion = form.cleaned_data['observacion']
                    limp = request.POST['id_limpiar']
                    # danio = request.POST['id_danio']
                    piezaparte = request.POST.getlist('piezaparte[]')
                    caracteristica = request.POST.getlist('catacteristica[]')
                    danio = request.POST.getlist('danio[]')
                    daniodes = request.POST.getlist('danioop[]')

                    mantenimientosactivos = MantenimientosActivosPreventivos.objects.get(id=id_activoman)
                    mantenimientosactivos.activofijo_id = id_cpu
                    mantenimientosactivos.tipoactivo_id = id_tipact
                    mantenimientosactivos.estusu = estusu
                    mantenimientosactivos.fecha = fecha
                    mantenimientosactivos.horamax = horamax
                    if eli_arc == 'on':
                        mantenimientosactivos.archivo = None
                    mantenimientosactivos.minutomax = minutomax
                    mantenimientosactivos.funcionarecibe = estfrec
                    mantenimientosactivos.funcionaentrega = estfent
                    mantenimientosactivos.marca = marca
                    mantenimientosactivos.modelo = modelo
                    mantenimientosactivos.sbequipo = bsugiere
                    mantenimientosactivos.descbaja = dsugiere
                    mantenimientosactivos.observaciones = observacion
                    mantenimientosactivos.save(request)
                    if 'archivo' in request.FILES:
                        newfile = request.FILES['archivo']
                        newfile._name = generar_nombre("evidencia_comunicado", newfile._name)
                        mantenimientosactivos.archivo = newfile
                        mantenimientosactivos.save(request)
                    TareasActivosPreventivos.objects.filter(mantenimiento=mantenimientosactivos).delete()
                    elementos = limp.split(',')
                    for elemento in elementos:
                        detalle = TareasActivosPreventivos(mantenimiento=mantenimientosactivos,
                                                           grupos_id=elemento)
                        detalle.save(request)
                    TareasActivosPreventivosDanios.objects.filter(mantenimiento=mantenimientosactivos).delete()
                    c = 0
                    while c < len(danio):
                        detalle = TareasActivosPreventivosDanios(mantenimiento=mantenimientosactivos,
                                                                 grupos_id=int(MantenimientoGruDanios.objects.get(
                                                                     descripcion=danio[c], grupocategoria_id=id_tipact,
                                                                     status=True).id),
                                                                 estadodanio=daniodes[c])
                        detalle.save(request)
                        c += 1
                    PiezaParteActivosPreventivos.objects.filter(mantenimiento=mantenimientosactivos).delete()
                    counter = 0
                    while counter < len(piezaparte):
                        detalle = PiezaParteActivosPreventivos(mantenimiento=mantenimientosactivos,
                                                               piezaparte_id=int(HdPiezaPartes.objects.get(
                                                                   descripcion=piezaparte[counter], status=True).id),
                                                               descripcion=caracteristica[counter])
                        detalle.save(request)
                        counter += 1
                    log(u'Editó mantenimiento preventivo Activo: %s' % mantenimientosactivos, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'addtarjeta':
            try:
                f = DetalleTarjetaControlForm(request.POST)
                if f.is_valid():
                    tarjeta = TarjetaControl.objects.get(pk=int(request.POST['id']))
                    detalle = DetalleMantenimiento(tarjeta=tarjeta,
                                                   mantenimientorealizar=f.cleaned_data['mantenimientorealizar'],
                                                   aplicagarantia=f.cleaned_data['aplicagarantia'],
                                                   observacion=f.cleaned_data['observacion'],
                                                   fechaentrega=f.cleaned_data['fechaentrega'],
                                                   estado=1,
                                                   taller=f.cleaned_data['taller'])
                    detalle.save(request)
                    log(u'Adicionó tarjeta: %s' % detalle, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'ingresodemantenimiento':
            try:
                f = DetalleTarjetaControlForm(request.POST)
                if f.is_valid():
                    detalle = DetalleMantenimiento.objects.get(pk=int(request.POST['id']))
                    detalle.costomanodeobra = f.cleaned_data['costomanodeobra']
                    detalle.mantenimientorealizado = f.cleaned_data['mantenimientorealizado']
                    detalle.aplicagarantia = f.cleaned_data['aplicagarantia']
                    detalle.manodeobra = f.cleaned_data['manodeobra']
                    detalle.repuestos = f.cleaned_data['repuestos']
                    detalle.costomanodereparacion = f.cleaned_data['costomanodereparacion']
                    detalle.facturamanodeobra = f.cleaned_data['facturamanodeobra']
                    detalle.facturareparacion = f.cleaned_data['facturareparacion']
                    detalle.fecharecepcion = f.cleaned_data['fecharecepcion']
                    detalle.observacion = f.cleaned_data['observacion']
                    detalle.estado = 2
                    detalle.save(request)
                    log(u'ingreso de mantenimiento : %s' % detalle, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'edittarjeta':
            try:
                f = DetalleTarjetaControlForm(request.POST)
                if f.is_valid():
                    detalle = DetalleMantenimiento.objects.get(pk=int(request.POST['id']))
                    detalle.fechaentrega = f.cleaned_data['fechaentrega']
                    detalle.mantenimientorealizar = f.cleaned_data['mantenimientorealizar']
                    detalle.taller = f.cleaned_data['taller']
                    detalle.observacion = f.cleaned_data['observacion']
                    detalle.estado = 2
                    detalle.save(request)
                    log(u'Editó tarjeta: %s' % detalle, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'addtraslado':
            try:
                f = TrasladoMantenimientoForm(request.POST)
                if f.is_valid():
                    traslado = TrasladoMantenimiento(departamentosolicita=f.cleaned_data['departamentosolicita'],
                                                     fecha=datetime.now(),
                                                     asistentelogistica=f.cleaned_data['asistentelogistica'],
                                                     usuariobienes=f.cleaned_data['usuariobienes'],
                                                     taller=f.cleaned_data['taller'],
                                                     administradorcontrato=f.cleaned_data['administradorcontrato'],
                                                     observacion=f.cleaned_data['observacion'])
                    traslado.save(request)
                    log(u'Adicionó traslado: %s' % traslado, request, "add")
                    cursor = connection.cursor()
                    cursor.execute(
                        "INSERT INTO sagest_detalletrasladomantenimiento (usuario_creacion_id, fecha_creacion, codigotraslado_id, activo_id, observacion, status, seleccionado) "
                        "SELECT %s, now(), %s, id, '', true, false from sagest_activofijo WHERE statusactivo=1 AND responsable_id=%s;",
                        [request.user.id, traslado.id, traslado.usuariobienes.id])
                    return JsonResponse({"result": "ok", "id": traslado.id})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'addbaja':
            try:
                f = BajaActivoForm(request.POST)
                if f.is_valid():
                    if f.cleaned_data['fechaoficio'] > f.cleaned_data['fecha']:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"La fecha de oficio no debe ser mayor a la fecha de baja"})
                    baja = BajaActivo(fecha=f.cleaned_data['fecha'],
                                      tiposolicitud=f.cleaned_data['tiposolicitud'],
                                      solicitante=f.cleaned_data['solicitante'],
                                      tipobaja=f.cleaned_data['tipobaja'],
                                      ubicacionbienentrega=f.cleaned_data['ubicacionbienentrega'],
                                      usuariobienentrega=f.cleaned_data['usuariobienentrega'],
                                      custodioentrega=f.cleaned_data['custodioentrega'],
                                      # usuariorecibe=f.cleaned_data['usuariorecibe'],
                                      oficio=f.cleaned_data['oficio'],
                                      # cargorecibe=f.cleaned_data['cargorecibe'],
                                      fechaoficio=f.cleaned_data['fechaoficio'],
                                      experto=f.cleaned_data['experto'],
                                      contadorper=f.cleaned_data['contadorper'],
                                      usuarioejecuta=f.cleaned_data['usuarioejecuta'],
                                      observacion=f.cleaned_data['observacion'])
                    baja.save(request)
                    log(u'Adicionó baja: %s' % baja, request, "add")
                    cursor = connection.cursor()
                    cursor.execute(
                        "INSERT INTO sagest_detallebajaactivo (usuario_creacion_id, fecha_creacion, codigobaja_id, activo_id, status, seleccionado) "
                        "SELECT %s, now(), %s, id, True, False from sagest_activofijo WHERE statusactivo=1 AND responsable_id=%s AND ubicacion_id=%s AND custodio_id=%s;",
                        [request.user.id, baja.id, baja.usuariobienentrega.id, baja.ubicacionbienentrega.id,
                         baja.custodioentrega.id])
                    return JsonResponse({"result": "ok", "id": baja.id})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'addtipobaja':
            try:
                f = TipoBajaForm(request.POST)
                if f.is_valid():
                    if TipoBaja.objects.filter(nombre=f.cleaned_data['nombre']).exists():
                        return JsonResponse({"result": "bad", "mensaje": u"Registro ya existe"})
                    tipobaja = TipoBaja(nombre=f.cleaned_data['nombre'])
                    tipobaja.save(request)
                    log(u'Adicionó Tipo baja: %s' % tipobaja, request, "add")
                    return JsonResponse({"result": "ok", "id": tipobaja.id})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'edittraslado':
            try:
                f = TrasladoMantenimientoForm(request.POST)
                if f.is_valid():
                    traslado = TrasladoMantenimiento.objects.get(pk=int(request.POST['id']))
                    traslado.observacion = f.cleaned_data['observacion']
                    traslado.save(request)
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'finalizarcons':
            try:
                constatacion = ConstatacionFisica.objects.get(pk=int(request.POST['id']))
                constatacion.estado = 2
                constatacion.fechafin = datetime.now()
                constatacion.save(request)
                log(u'Finalizar constatacion: %s' % constatacion, request, "edit")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'confirmarexpo':
            try:
                exportacion = ExportacionesActivos.objects.get(pk=int(request.POST['id']))
                exportacion.estado = 2
                exportacion.save(request)
                lista = exportacion.activos.all().values_list('id', flat=True)
                ActivoFijo.objects.filter(id__in=lista).update(subidogobierno=True)
                log(u'Finalizar exportacion: %s' % exportacion, request, "edit")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'finalizartraslado':
            try:
                traslado = TrasladoMantenimiento.objects.get(pk=int(request.POST['id']))
                traslado.estado = 2
                traslado.save(request)
                secuencia = secuencia_activos(request)
                if not traslado.numero:
                    secuencia.numeromantenimiento += 1
                    secuencia.save(request)
                    traslado.numero = secuencia.numeromantenimiento
                    traslado.save(request)
                traslado.detalletrasladomantenimiento_set.filter(seleccionado=False).update(status=False)
                for detalle in traslado.detalletrasladomantenimiento_set.filter(seleccionado=True):
                    detallematenimiento = DetalleMantenimiento(tarjeta=detalle.activo.mi_tarjeta_control(),
                                                               detalletrasladomantenimiento=detalle,
                                                               taller=traslado.taller,
                                                               fechaentrega=traslado.fecha,
                                                               mantenimientorealizar=detalle.observacion)
                    detallematenimiento.save(request)
                log(u'Finalizar traslado: %s' % traslado, request, "edit")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'finalizartraspaso':
            try:
                responsable = None
                solicitud = None
                lista = []
                traspaso = TraspasoActivo.objects.get(pk=int(request.POST['id']))
                if not traspaso.detalletraspasoactivo_set.filter(seleccionado=True).exists():
                    return JsonResponse({"result": "bad", "mensaje": u"Debe seleccionar un activo."})
                traspaso.estado = 2
                traspaso.save(request)
                secuencia = secuencia_activos(request)
                if not traspaso.numero:
                    secuencia.numerotraspaso += 1
                    secuencia.save(request)
                    traspaso.numero = secuencia.numerotraspaso
                traspaso.normativatraspaso = secuencia.normativatraspaso
                traspaso.save(request)
                traspaso.detalletraspasoactivo_set.filter(seleccionado=False).update(status=False)
                traspaso.totalbienes = traspaso.cantidad_seleccionados()
                traspaso.save(request)
                data[
                    'ubicacionentrega'] = traspaso.ubicacionbienentrega.nombre if traspaso.ubicacionbienentrega.nombre else ''
                data[
                    'bloqueentrega'] = traspaso.ubicacionbienentrega.bloque.descripcion if traspaso.ubicacionbienentrega.bloque.descripcion else ''
                data[
                    'bloquerecibe'] = traspaso.ubicacionbienrecibe.bloque.descripcion if traspaso.ubicacionbienrecibe.bloque.descripcion else ''
                for detalle in traspaso.detalletraspasoactivo_set.filter(status=True):
                    activo = detalle.activo
                    responsable = traspaso.usuariobienrecibe if traspaso.usuariobienrecibe else activo.responsable
                    responsableanterior = activo.responsable
                    lista.append(detalle.activo.id)
                    activo.actualiza_responsable_directo(traspaso.custodiobienrecibe, traspaso.ubicacionbienrecibe,
                                                         traspaso.usuariobienrecibe if traspaso.usuariobienrecibe else activo.responsable)
                    solicitud = SolicitudActivos.objects.filter(status=True, activo_id=activo.pk,
                                                                responsableasignacion=responsable,estado=2)
                    if solicitud:
                        historial = HistorialTraspaso(activofijo_id=activo.id, responsableanterior=responsableanterior,
                                                      responsableasignado=responsable,
                                                      fechasolicitud=solicitud[0].fechasolicitud,
                                                      fechaasignacion=datetime.now().date())
                        solicitud[0].aceptado = 1
                        solicitud[0].save(request)
                        historial.save(request)
                        log(u'Solicitud traspaso confirmada: %s' % solicitud[0], request, "add")
                    activo.save(request)
                log(u'Finalizar traspaso: %s' % traspaso, request, "edit")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse(
                    {"result": "bad", "mensaje": '{} Error on line {}'.format(ex, sys.exc_info()[-1].tb_lineno)})

        elif action == 'abrirtraspaso':
            try:
                traspaso = TraspasoActivo.objects.get(pk=int(request.POST['id']))
                traspaso.estado = 1
                traspaso.save(request)
                traspaso.detalletraspasoactivo_set.all().update(status=True)
                for detalle in traspaso.detalletraspasoactivo_set.all():
                    activo = detalle.activo
                    activo.actualiza_responsable()
                    activo.save(request)
                log(u'Abrió traspaso: %s' % traspaso, request, "edit")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'actualizamasivo':
            try:
                activos = ActivoFijo.objects.filter(responsable__isnull=True, status=True)
                for act in activos:
                    traspaso = act.detalletraspasoactivo_set.filter(status=True).order_by(
                        '-codigotraspaso__fecha').last()
                    act.custodio = traspaso.codigotraspaso.custodiobienrecibe
                    act.ubicacion = traspaso.codigotraspaso.ubicacionbienrecibe
                    act.responsable = traspaso.codigotraspaso.usuariobienrecibe
                    act.save(request)
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'eliminartraspaso':
            try:
                traspaso = TraspasoActivo.objects.get(pk=int(request.POST['id']))
                traspaso.status = False
                traspaso.save(request)
                traspaso.detalletraspasoactivo_set.all().update(status=False)
                log(u'Eliminar traspaso: %s' % traspaso, request, "delete")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'deleteexpo':
            try:
                exportacion = ExportacionesActivos.objects.get(pk=int(request.POST['id']))
                exportacion.delete()
                log(u'Eliminar exportacion: %s' % exportacion, request, "delete")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'deleteimpo':
            try:
                exportacion = ArchivoActivoFijo.objects.get(pk=int(request.POST['id']))
                exportacion.delete()
                log(u'Eliminar importacion: %s' % exportacion, request, "delete")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})
        elif action == 'delete':
            try:
                with transaction.atomic():
                    instancia = CatalogoBien.objects.get(pk=int(request.POST['id']))
                    activotecnologico = ActivoTecnologico.objects.filter(status=True, activotecnologico__catalogo_id=instancia.id)
                    for activoat in activotecnologico:
                        activoat.status = False
                        activoat.save()
                        log(u'Elimino activo tecnologico: %s' % activoat, request, "delete")
                    instancia.clasificado = False
                    instancia.equipoelectronico = False
                    instancia.save(request)
                    log(u'Elimino Catalogo: %s' % instancia, request, "delete")
                    res_json = {"error": False}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)
        elif action == 'eliminartraslado':
            try:
                traslado = TrasladoMantenimiento.objects.get(pk=int(request.POST['id']))
                traslado.status = False
                traslado.save(request)
                traslado.detalletrasladomantenimiento_set.all().update(status=False)
                log(u'Eliminar traspaso: %s' % traslado, request, "delete")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'eliminarconstatacion':
            try:
                constatacion = ConstatacionFisica.objects.get(pk=int(request.POST['id']))
                constatacion.status = False
                constatacion.save(request)
                constatacion.detalleconstatacionfisica_set.all().update(status=False)
                constatacion.detallenoidentificado_set.all().update(status=False)
                log(u'Eliminar traspaso: %s' % constatacion, request, "delete")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'eliminarbaja':
            try:
                baja = BajaActivo.objects.get(pk=int(request.POST['id']))
                baja.status = False
                baja.save(request)
                baja.detallebajaactivo_set.all().update(status=False)
                log(u'Eliminar baja: %s' % baja, request, "delete")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'eliminartipobaja':
            try:
                tipobaja = TipoBaja.objects.get(pk=int(request.POST['id']))
                log(u'Eliminar Tipo baja: %s' % tipobaja, request, "delete")
                tipobaja.delete()
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'finalizarbaja':
            try:
                baja = BajaActivo.objects.get(pk=int(request.POST['id']))
                baja.estado = 2
                baja.save(request)
                secuencia = secuencia_activos(request)
                if not baja.numero:
                    secuencia.numerobajaactivo += 1
                    secuencia.save(request)
                    baja.numero = secuencia.numerobajaactivo
                    baja.save(request)
                baja.detallebajaactivo_set.filter(seleccionado=False).update(status=False)
                for detalle in baja.detallebajaactivo_set.filter(seleccionado=True):
                    activo = detalle.activo
                    activo.depreciar_activo(baja.fecha, request)
                    activo.statusactivo = 2
                    activo.save(request)
                log(u'Finalizar traslado: %s' % baja, request, "edit")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'finalizarmantenimiento':
            try:
                mantenimiento = DetalleTrasladoMantenimiento.objects.get(pk=int(request.POST['id']))
                detalle = mantenimiento.detallemantenimiento_set.all()[0]
                detalle.estado = 2
                detalle.save(request)
                log(u'Finalizar tarjeta de control: %s' % detalle, request, "edit")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'finalizaracta':
            try:
                acta = TraspasoActivo.objects.get(pk=int(request.POST['id']))
                acta.estado = 2
                acta.save(request)
                log(u'Finalizar acta de entrega: %s' % acta, request, "edit")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'cambiousuario':
            try:
                f = ConstatacionForm(request.POST)
                if f.is_valid():
                    usuario = Persona.objects.get(id=int(f.cleaned_data['usuariobienes']))
                    constatacion = ConstatacionFisica.objects.get(pk=int(request.POST['id']))
                    constatacion.detalleconstatacionfisica_set.all().delete()
                    constatacion.usuariobienes = usuario
                    constatacion.ubicacionbienes = f.cleaned_data['ubicacionbienes']
                    constatacion.save(request)
                    cursor = connection.cursor()
                    cursor.execute(
                        "INSERT INTO sagest_detalleconstatacionfisica (status, usuario_creacion_id, fecha_creacion, codigoconstatacion_id, activo_id, perteneceusuario, requieretraspaso, estadooriginal_id, estadoactual_id, encontrado, enuso, observacion) "
                        "SELECT true, %s, now(), %s, id, true, false, estado_id, estado_id, false, false, ''  from sagest_activofijo WHERE statusactivo=1 AND responsable_id=%s AND ubicacion_id=%s;",
                        [request.user.id, constatacion.id, constatacion.usuariobienes.id,
                         constatacion.ubicacionbienes.id])
                    return JsonResponse({"result": "ok", "id": constatacion.id})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'addconstatacion':
            try:
                f = ConstatacionForm(request.POST)
                if f.is_valid():
                    secuencia = secuencia_activos(request)
                    secuencia.numeroconstatacion += 1
                    secuencia.save(request)
                    constatacion = ConstatacionFisica(fechainicio=datetime.now(),
                                                      estado=1,
                                                      numero=secuencia.numeroconstatacion,
                                                      normativaconstatacion=secuencia.normativaconstatacion,
                                                      usuariobienes_id=f.cleaned_data['usuariobienes'],
                                                      ubicacionbienes=f.cleaned_data['ubicacionbienes'],
                                                      observacion=f.cleaned_data['observacion'])
                    constatacion.save(request)
                    log(u'Adicionó contratación Fisica: %s' % constatacion, request, "add")
                    cursor = connection.cursor()
                    cursor.execute(
                        "INSERT INTO sagest_detalleconstatacionfisica (usuario_creacion_id, fecha_creacion, codigoconstatacion_id, activo_id, perteneceusuario, requieretraspaso, estadooriginal_id, estadoactual_id, encontrado, enuso, status, observacion) "
                        "SELECT %s, now(), %s, id, true, false, estado_id, estado_id, false, false, true, '' from sagest_activofijo WHERE statusactivo=1 AND responsable_id=%s AND ubicacion_id=%s;",
                        [request.user.id, constatacion.id, constatacion.usuariobienes.id,
                         constatacion.ubicacionbienes.id])
                    return JsonResponse({"result": "ok", "id": constatacion.id})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'addtraspaso':
            try:
                f = TraspasoActivoForm(request.POST)
                if f.is_valid():
                    entrega = f.cleaned_data['usuariobienentrega']
                    if f.cleaned_data['fechaoficio'] > f.cleaned_data['fecha']:
                        return JsonResponse({"result": "bad",
                                             "mensaje": u"La fecha de oficio no debe ser mayor a la fecha de traspaso"})
                    if f.cleaned_data['custodiobienrecibe'] == f.cleaned_data['custodiobienentrega'] and f.cleaned_data[
                        'usuariobienentrega'] == f.cleaned_data['usuariobienrecibe'] and f.cleaned_data[
                        'ubicacionbienentrega'] == f.cleaned_data['ubicacionbienrecibe']:
                        return JsonResponse({"result": "bad",
                                             "mensaje": u"Está tratando de realizar un traspasos. con datos de entrega y recepción similares"})
                    traspaso = TraspasoActivo(custodiobienrecibe=f.cleaned_data['custodiobienrecibe'],
                                              usuariobienentrega=f.cleaned_data['usuariobienentrega'],
                                              custodiobienentrega=f.cleaned_data['custodiobienentrega'],
                                              usuariobienrecibe=f.cleaned_data['usuariobienrecibe'],
                                              ubicacionbienentrega=f.cleaned_data['ubicacionbienentrega'],
                                              responsablebienes_id=1204,
                                              ubicacionbienrecibe=f.cleaned_data['ubicacionbienrecibe'],
                                              fecha=f.cleaned_data['fecha'],
                                              solicitante=f.cleaned_data['solicitante'],
                                              fechaoficio=f.cleaned_data['fechaoficio'],
                                              observacion=f.cleaned_data['observacion'],
                                              oficio=f.cleaned_data['oficio'],
                                              tiposolicitud=f.cleaned_data['tiposolicitud'],
                                              tipotraspaso=1,
                                              tipo=2)
                    traspaso.save(request)
                    log(u'Adicionó traspaso activo: %s' % traspaso, request, "add")
                    cursor = connection.cursor()
                    cursor.execute(
                        "INSERT INTO sagest_detalletraspasoactivo (usuario_creacion_id, fecha_creacion, codigotraspaso_id, activo_id, status, historico, seleccionado) "
                        "SELECT %s, now(), %s, id, true, false, false from sagest_activofijo WHERE statusactivo=1 AND responsable_id=%s AND ubicacion_id=%s AND custodio_id=%s;",
                        [request.user.id, traspaso.id, traspaso.usuariobienentrega.id, traspaso.ubicacionbienentrega.id,
                         traspaso.custodiobienentrega.id])
                    return JsonResponse({"result": "ok", "id": traspaso.id})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'addtraspasocustodio':
            try:
                f = TraspasoActivoCustodioForm(request.POST)
                if f.is_valid():
                    if f.cleaned_data['fechaoficio'] > f.cleaned_data['fecha']:
                        return JsonResponse({"result": "bad",
                                             "mensaje": u"La fecha de oficio no debe ser mayor a la fecha de traspaso"})
                    if f.cleaned_data['custodiobienrecibe'] == f.cleaned_data['custodiobienentrega']:
                        return JsonResponse({"result": "bad",
                                             "mensaje": u"Está tratando de realizar un traspasos. con datos de entrega y recepción similares"})
                    traspaso = TraspasoActivo(custodiobienrecibe=f.cleaned_data['custodiobienrecibe'],
                                              custodiobienentrega=f.cleaned_data['custodiobienentrega'],
                                              ubicacionbienentrega=f.cleaned_data['ubicacionbienentrega'],
                                              ubicacionbienrecibe=f.cleaned_data['ubicacionbienentrega'],
                                              responsablebienes_id=1204,
                                              fecha=f.cleaned_data['fecha'],
                                              solicitante=f.cleaned_data['solicitante'],
                                              fechaoficio=f.cleaned_data['fechaoficio'],
                                              observacion=f.cleaned_data['observacion'],
                                              oficio=f.cleaned_data['oficio'],
                                              tiposolicitud=f.cleaned_data['tiposolicitud'],
                                              tipotraspaso=2,
                                              tipo=2)
                    traspaso.save(request)
                    log(u'Adicionó traspaso custio de  activo: %s' % traspaso, request, "add")
                    cursor = connection.cursor()
                    cursor.execute(
                        "INSERT INTO sagest_detalletraspasoactivo (usuario_creacion_id, fecha_creacion, codigotraspaso_id, activo_id, historico, seleccionado, status) "
                        "SELECT %s, now(), %s, id, false, false, true from sagest_activofijo WHERE statusactivo=1 AND custodio_id=%s AND ubicacion_id=%s;",
                        [request.user.id, traspaso.id, traspaso.custodiobienentrega.id,
                         traspaso.ubicacionbienentrega.id])
                    return JsonResponse({"result": "ok", "id": traspaso.id})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'adddirector':
            try:
                f = DirectorResponsableBajaForm(request.POST)
                if f.is_valid():
                    if f.cleaned_data['fechafin']:
                        if f.cleaned_data['fechainicio'] > f.cleaned_data['fechafin']:
                            return JsonResponse(
                                {"result": "bad", "mensaje": u"La fecha inicio no puede ser mayor que la fecha fin "})

                    director = DirectorResponsableBaja(
                        responsable_id=f.cleaned_data['responsable'],
                        fechainicio=f.cleaned_data['fechainicio'],
                    )
                    if f.cleaned_data['actual']:
                        director.actual = True
                        if DirectorResponsableBaja.objects.filter(status=True, actual=True).exists():
                            cerrar = DirectorResponsableBaja.objects.get(status=True, actual=True)
                            cerrar.fechafin = datetime.now().date()
                            cerrar.actual = False
                            cerrar.save()
                    else:
                        director.fechafin = f.cleaned_data['fechafin']
                    director.save(request)
                    log(u'Adicionó director: %s' % director, request, "add")
                    return JsonResponse({"result": False}, safe=False)
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'editdirector':
            try:
                f = DirectorResponsableBajaForm(request.POST)
                director = DirectorResponsableBaja.objects.get(pk=int(request.POST['id']))
                if f.is_valid():
                    if f.cleaned_data['fechafin']:
                        if f.cleaned_data['fechainicio'] > f.cleaned_data['fechafin']:
                            return JsonResponse(
                                {"result": "bad", "mensaje": u"La fecha inicio no puede ser mayor que la fecha fin "})

                    director.fechainicio = f.cleaned_data['fechainicio']
                    if f.cleaned_data['actual']:
                        director.actual = True
                        if DirectorResponsableBaja.objects.filter(status=True, actual=True).exists():
                            cerrar = DirectorResponsableBaja.objects.get(status=True, actual=True)
                            cerrar.fechafin = datetime.now().date()
                            cerrar.actual = False
                            cerrar.save()
                    else:
                        director.fechafin = f.cleaned_data['fechafin']
                    director.save(request)
                    log(u'Adicionó director: %s' % director, request, "add")
                    return JsonResponse({"result": False}, safe=False)
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'activosusuarioconstatacion':
            try:
                data = {}
                data['constatacion'] = constatacion = ConstatacionFisica.objects.get(pk=int(request.POST['id']))
                detalle = constatacion.detalleconstatacionfisica_set.filter(perteneceusuario=True)
                pagina = 1
                paging = MiPaginador(detalle, 10)
                p = 1
                try:
                    if 'page' in request.POST:
                        p = int(request.POST['page'])
                    page = paging.page(p)
                except:
                    page = paging.page(1)
                data['paging'] = paging
                data['rangospaging'] = paging.rangos_paginado(p)
                data['page'] = page
                data['detalles'] = page.object_list
                data['usuario'] = request.user
                template = get_template("inventario_activofijo/detalleconstatacion.html")
                json_content = template.render(data)
                return JsonResponse(
                    {"result": "ok", 'data': json_content, 'contador': constatacion.contadores_pertenece()})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'activostraspaso':
            try:
                data = {}
                data['traspaso'] = traspaso = TraspasoActivo.objects.get(pk=int(request.POST['id']))
                detalle = traspaso.detalletraspasoactivo_set.all()
                pagina = 1
                paging = MiPaginador(detalle, 10)
                p = 1
                try:
                    if 'page' in request.POST:
                        p = int(request.POST['page'])
                    page = paging.page(p)
                except:
                    page = paging.page(1)
                data['paging'] = paging
                data['rangospaging'] = paging.rangos_paginado(p)
                data['page'] = page
                data['detalles'] = page.object_list
                data['usuario'] = request.user
                template = get_template("inventario_activofijo/detalletraspaso.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'data': json_content, 'contador': traspaso.contador_traspaso()})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'activosacta':
            try:
                data = {}
                data['acta'] = acta = TraspasoActivo.objects.get(pk=int(request.POST['id']))
                detalle = acta.detalletraspasoactivo_set.all()
                pagina = 1
                paging = MiPaginador(detalle, 10)
                p = 1
                try:
                    if 'page' in request.POST:
                        p = int(request.POST['page'])
                    page = paging.page(p)
                except:
                    page = paging.page(1)
                data['paging'] = paging
                data['rangospaging'] = paging.rangos_paginado(p)
                data['page'] = page
                data['detalles'] = page.object_list
                data['usuario'] = request.user
                template = get_template("inventario_activofijo/detalleacta.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'data': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'activosbaja':
            try:
                data = {}
                data['baja'] = baja = BajaActivo.objects.get(pk=int(request.POST['id']))
                detalle = baja.detallebajaactivo_set.all()
                pagina = 1
                paging = MiPaginador(detalle, 10)
                p = 1
                try:
                    if 'page' in request.POST:
                        p = int(request.POST['page'])
                    page = paging.page(p)
                except:
                    page = paging.page(1)
                data['paging'] = paging
                data['rangospaging'] = paging.rangos_paginado(p)
                data['page'] = page
                data['detalles'] = page.object_list
                data['usuario'] = request.user
                template = get_template("inventario_activofijo/detallebajaactivos.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'data': json_content, 'contador': baja.contador()})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'activostraslado':
            try:
                data = {}
                data['traslado'] = traslado = TrasladoMantenimiento.objects.get(pk=int(request.POST['id']))
                detalle = traslado.detalletrasladomantenimiento_set.all()
                pagina = 1
                paging = MiPaginador(detalle, 10)
                p = 1
                try:
                    if 'page' in request.POST:
                        p = int(request.POST['page'])
                    page = paging.page(p)
                except:
                    page = paging.page(1)
                data['paging'] = paging
                data['rangospaging'] = paging.rangos_paginado(p)
                data['page'] = page
                data['detalles'] = page.object_list
                data['usuario'] = request.user
                template = get_template("inventario_activofijo/activostraslado.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'data': json_content, 'contador': traslado.contador()})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'activosotrousuarioconstatacion':
            try:
                data = {}
                constatacion = ConstatacionFisica.objects.get(pk=int(request.POST['id']))
                detalle = constatacion.detalleconstatacionfisica_set.filter(perteneceusuario=False).order_by(
                    'activo__codigointerno', 'activo__codigogobierno')
                pagina = 1
                paging = MiPaginador(detalle, 10)
                p = 1
                try:
                    if 'page' in request.POST:
                        p = int(request.POST['page'])
                    page = paging.page(p)
                except:
                    page = paging.page(1)
                data['paging'] = paging
                data['rangospaging'] = paging.rangos_paginado(p)
                data['page'] = page
                data['detalles'] = page.object_list
                data['usuario'] = request.user
                template = get_template("inventario_activofijo/detalleotrosconstatacion.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'data': json_content, 'contador': constatacion.contadores_otros()})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'activosnoidentificadosconstatacion':
            try:
                data = {}
                constatacion = ConstatacionFisica.objects.get(pk=int(request.POST['id']))
                detalle = constatacion.detallenoidentificado_set.all().order_by('codigobarra')
                pagina = 1
                paging = MiPaginador(detalle, 10)
                p = 1
                try:
                    if 'page' in request.POST:
                        p = int(request.POST['page'])
                    page = paging.page(p)
                except:
                    page = paging.page(1)
                data['paging'] = paging
                data['rangospaging'] = paging.rangos_paginado(p)
                data['page'] = page
                data['detalles'] = page.object_list
                data['usuario'] = request.user
                template = get_template("inventario_activofijo/detalleniconstatacion.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'data': json_content, 'contador': constatacion.contadores_ni()})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'marcarencontrado':
            try:
                detalle = DetalleConstatacionFisica.objects.get(pk=int(request.POST['id']))
                encontrado = request.POST['valencontrado'] == u'true'
                if encontrado:
                    detalle.encontrado = True
                    detalle.enuso = True
                else:
                    detalle.encontrado = False
                    detalle.enuso = False
                    detalle.requieretraspaso = False
                detalle.save(request)
                return JsonResponse(
                    {"result": "ok", "reload": 'False', 'contador': detalle.codigoconstatacion.contadores_pertenece()})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos"})

        elif action == 'marcartodosencontrado':
            try:
                constatacion = ConstatacionFisica.objects.get(pk=int(request.POST['id']))
                if constatacion.detalleconstatacionfisica_set.all()[0].encontrado == True:
                    constatacion.detalleconstatacionfisica_set.filter(
                        Q(usuario_modificacion__isnull=True) | Q(usuario_modificacion=request.user)).update(
                        encontrado=False, enuso=False, requieretraspaso=False)
                else:
                    constatacion.detalleconstatacionfisica_set.filter(
                        Q(usuario_modificacion__isnull=True) | Q(usuario_modificacion=request.user)).update(
                        encontrado=True, enuso=True)
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos"})

        elif action == 'marcarenuso':
            try:
                detalle = DetalleConstatacionFisica.objects.get(pk=int(request.POST['id']))
                detalle.enuso = request.POST['valenuso'] == u'true'
                detalle.save(request)
                return JsonResponse(
                    {"result": "ok", "reload": 'False', 'contador': detalle.codigoconstatacion.contadores_pertenece()})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos"})

        elif action == 'marcarenusootro':
            try:
                detalle = DetalleConstatacionFisica.objects.get(pk=int(request.POST['id']))
                detalle.enuso = request.POST['valenuso'] == u'true'
                detalle.save(request)
                return JsonResponse(
                    {"result": "ok", "reload": 'False', 'contador': detalle.codigoconstatacion.contadores_otros()})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos"})

        elif action == 'marcartodostraslado':
            try:
                traslado = TrasladoMantenimiento.objects.get(pk=int(request.POST['id']))
                estado = request.POST['valor'] == u'true'
                traslado.detalletrasladomantenimiento_set.filter(
                    Q(usuario_modificacion=request.user) | Q(usuario_modificacion=None)).update(seleccionado=estado,
                                                                                                usuario_modificacion=request.user,
                                                                                                fecha_modificacion=datetime.now())
                return JsonResponse({"result": "ok", "reload": 'True', 'contador': traslado.contador()})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos"})

        elif action == 'marcartodostraspaso':
            try:
                traspaso = TraspasoActivo.objects.get(pk=int(request.POST['id']))
                estado = request.POST['valor'] == u'true'
                traspaso.detalletraspasoactivo_set.all().update(seleccionado=estado)
                return JsonResponse({"result": "ok", "reload": 'True', 'contador': traspaso.contador_traspaso()})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos"})

        elif action == 'marcartodosbaja':
            try:
                baja = BajaActivo.objects.get(pk=int(request.POST['id']))
                estado = request.POST['valor'] == u'true'
                baja.detallebajaactivo_set.all().update(seleccionado=estado)
                return JsonResponse({"result": "ok", "reload": 'True', 'contador': baja.contador()})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos"})

        elif action == 'marcarseleccionadotraslado':
            try:
                detalle = DetalleTrasladoMantenimiento.objects.get(pk=int(request.POST['id']))
                detalle.seleccionado = request.POST['valor'] == u'true'
                detalle.save(request)
                return JsonResponse({"result": "ok", "reload": 'False', 'contador': detalle.codigotraslado.contador()})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos"})

        elif action == 'marcarseleccionadotraspaso':
            try:
                detalle = DetalleTraspasoActivo.objects.get(pk=int(request.POST['id']))
                detalle.seleccionado = request.POST['valor'] == u'true'
                detalle.save(request)
                traspaso = detalle.codigotraspaso
                traspaso.totalbienes = traspaso.cantidad_seleccionados()
                return JsonResponse(
                    {"result": "ok", "reload": 'False', 'contador': detalle.codigotraspaso.contador_traspaso()})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos"})

        elif action == 'marcarseleccionadobaja':
            try:
                detalle = DetalleBajaActivo.objects.get(pk=int(request.POST['id']))
                detalle.seleccionado = request.POST['valor'] == u'true'
                detalle.save(request)
                return JsonResponse({"result": "ok", "reload": 'False', 'contador': detalle.codigobaja.contador()})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos"})

        elif action == 'actualizaobservaciontraslado':
            try:
                detalle = DetalleTrasladoMantenimiento.objects.get(pk=int(request.POST['id']))
                detalle.observacion = request.POST['valor']
                detalle.save(request)
                return JsonResponse({"result": "ok", "reload": 'False', 'contador': detalle.codigotraslado.contador()})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos"})

        elif action == 'eliminar_otro':
            try:
                detalle = DetalleConstatacionFisica.objects.get(pk=int(request.POST['id']))
                detalle.delete()
                return JsonResponse(
                    {"result": "ok", "reload": 'False', 'contador': detalle.codigoconstatacion.contadores_ni()})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos"})

        elif action == 'eliminar_noiden':
            try:
                detalle = DetalleNoIdentificado.objects.get(pk=int(request.POST['id']))
                detalle.delete()
                return JsonResponse(
                    {"result": "ok", "reload": 'False', 'contador': detalle.codigoconstatacion.contadores_otros()})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos"})

        elif action == 'marcarrequieretrasotro':
            try:
                detalle = DetalleConstatacionFisica.objects.get(pk=int(request.POST['id']))
                detalle.requieretraspaso = request.POST['valrequieretrasotro'] == u'true'
                detalle.save(request)
                return JsonResponse(
                    {"result": "ok", "reload": 'False', 'contador': detalle.codigoconstatacion.contadores_otros()})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos"})

        elif action == 'marcartras':
            try:
                detalle = DetalleConstatacionFisica.objects.get(pk=int(request.POST['id']))
                encontrado = request.POST['valrequieretras'] == u'true'
                if encontrado:
                    detalle.requieretraspaso = True
                else:
                    detalle.requieretraspaso = False
                detalle.save(request)
                return JsonResponse(
                    {"result": "ok", "reload": 'False', 'contador': detalle.codigoconstatacion.contadores_pertenece()})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos"})

        elif action == 'buscarcodigoconstatacion':
            try:
                constatacion = ConstatacionFisica.objects.get(pk=int(request.POST['id']))
                if constatacion.detalleconstatacionfisica_set.filter(
                        Q(activo__codigogobierno=request.POST['codigo']) | Q(
                                activo__codigointerno=request.POST['codigo']), perteneceusuario=True).exists():
                    detalle = constatacion.detalleconstatacionfisica_set.filter(
                        Q(activo__codigogobierno=request.POST['codigo']) | Q(
                            activo__codigointerno=request.POST['codigo']), perteneceusuario=True)[0]
                    detallecons = constatacion.detalleconstatacionfisica_set.filter(perteneceusuario=True).order_by(
                        'activo__codigointerno', 'activo__codigogobierno')
                    pagina = ((list(detallecons.values_list('id', flat=True)).index(detalle.id) + 1) / 10.0)
                    try:
                        rest = 0
                        if len(str(pagina).split('.')) > 1:
                            rest = (1 if int(str(pagina).split('.')[1]) > 0 else 0)
                        pagina = int(str(pagina).split('.')[0]) + rest
                    except Exception as ex:
                        pass
                    return JsonResponse(
                        {"result": "ok", "reload": 'True', 'encontrado': 'True', 'pertenece': 'True', "pagina": pagina,
                         "id": detalle.id})
                else:
                    if ActivoFijo.objects.filter(
                            Q(codigogobierno=request.POST['codigo']) | Q(codigointerno=request.POST['codigo']),
                            statusactivo=1).exists():
                        if constatacion.detalleconstatacionfisica_set.filter(
                                Q(activo__codigogobierno=request.POST['codigo']) | Q(
                                        activo__codigointerno=request.POST['codigo']), perteneceusuario=False).exists():
                            return JsonResponse({"result": "bad", "mensaje": u"Activo ya adicionado"})
                        activo = ActivoFijo.objects.filter(
                            Q(codigogobierno=request.POST['codigo']) | Q(codigointerno=request.POST['codigo']),
                            statusactivo=1).distinct()[0]
                        detalle = DetalleConstatacionFisica(codigoconstatacion=constatacion,
                                                            perteneceusuario=False,
                                                            activo=activo,
                                                            usuariobienes=activo.responsable,
                                                            ubicacionbienes=activo.ubicacion,
                                                            estadooriginal=activo.estado,
                                                            estadoactual=activo.estado,
                                                            encontrado=True)
                        detalle.save(request)
                        detallecons = constatacion.detalleconstatacionfisica_set.filter(
                            perteneceusuario=False).order_by('activo__codigointerno', 'activo__codigogobierno')
                        pagina = ((list(detallecons.values_list('id', flat=True)).index(detalle.id) + 1) / 10.0)
                        try:
                            rest = 0
                            if len(str(pagina).split('.')) > 1:
                                rest = (1 if int(str(pagina).split('.')[1]) > 0 else 0)
                            pagina = int(str(pagina).split('.')[0]) + rest
                        except:
                            pass
                        return JsonResponse(
                            {"result": "ok", "reload": 'True', 'encontrado': 'True', 'pertenece': 'False',
                             "pagina": pagina, "id": detalle.id})
                    else:
                        if ActivoFijo.objects.filter(
                                Q(codigogobierno=request.POST['codigo']) | Q(codigointerno=request.POST['codigo']),
                                statusactivo=2).exists():
                            return JsonResponse({"result": "bad", "mensaje": u"El Activo se encuentra dado de baja"})
                        if constatacion.detallenoidentificado_set.filter(codigobarra=request.POST['codigo']).exists():
                            return JsonResponse(
                                {"result": "bad", "mensaje": u"Ya se encuentra registrado como No Identificado"})
                        return JsonResponse({"result": "ok", 'encontrado': 'False', 'pertenece': 'False'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al buscar los datos"})

        elif action == 'buscarcodigotraslado':
            try:
                traslado = TrasladoMantenimiento.objects.get(pk=int(request.POST['id']))
                if not traslado.detalletrasladomantenimiento_set.filter(
                        Q(activo__codigogobierno=request.POST['codigo']) | Q(
                                activo__codigointerno=request.POST['codigo'])).exists():
                    return JsonResponse({"result": "bad", "mensaje": u"El activo no se encuentra en lista"})
                detalle = traslado.detalletrasladomantenimiento_set.filter(
                    Q(activo__codigogobierno=request.POST['codigo']) | Q(activo__codigointerno=request.POST['codigo']))[
                    0]
                detalletrasl = traslado.detalletrasladomantenimiento_set.all().distinct().order_by(
                    'activo__codigointerno', 'activo__codigogobierno')
                pagina = ((list(detalletrasl.values_list('id', flat=True)).index(detalle.id) + 1) / 10.0)
                try:
                    rest = 0
                    if len(str(pagina).split('.')) > 1:
                        rest = (1 if int(str(pagina).split('.')[1]) > 0 else 0)
                    pagina = int(str(pagina).split('.')[0]) + rest
                except:
                    pass
                return JsonResponse({"result": "ok", "encontrado": True, "pagina": pagina, "id": detalle.id})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al buscar los datos"})

        elif action == 'buscarcodigotraspaso':
            try:
                traspaso = TraspasoActivo.objects.get(pk=int(request.POST['id']))
                if not traspaso.detalletraspasoactivo_set.filter(Q(activo__codigogobierno=request.POST['codigo']) | Q(
                        activo__codigointerno=request.POST['codigo'])).exists():
                    return JsonResponse({"result": "bad", "mensaje": u"El activo no se encuentra en lista"})
                detalle = traspaso.detalletraspasoactivo_set.filter(
                    Q(activo__codigogobierno=request.POST['codigo']) | Q(activo__codigointerno=request.POST['codigo']))[
                    0]
                detalletrasp = traspaso.detalletraspasoactivo_set.all().distinct().order_by('activo__codigointerno',
                                                                                            'activo__codigogobierno')
                pagina = ((list(detalletrasp.values_list('id', flat=True)).index(detalle.id) + 1) / 10.0)
                try:
                    rest = 0
                    if len(str(pagina).split('.')) > 1:
                        rest = (1 if int(str(pagina).split('.')[1]) > 0 else 0)
                    pagina = int(str(pagina).split('.')[0]) + rest
                except:
                    pass
                return JsonResponse({"result": "ok", "encontrado": True, "pagina": pagina, "id": detalle.id})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al buscar los datos"})

        elif action == 'buscarcodigobaja':
            try:
                baja = BajaActivo.objects.get(pk=int(request.POST['id']))
                if not baja.detallebajaactivo_set.filter(Q(activo__codigogobierno=request.POST['codigo']) | Q(
                        activo__codigointerno=request.POST['codigo'])).exists():
                    return JsonResponse({"result": "bad", "mensaje": u"El activo no se encuentra en lista"})
                detalle = baja.detallebajaactivo_set.filter(
                    Q(activo__codigogobierno=request.POST['codigo']) | Q(activo__codigointerno=request.POST['codigo']))[
                    0]
                detalletrasp = baja.detallebajaactivo_set.all().distinct().order_by('activo__codigointerno',
                                                                                    'activo__codigogobierno')
                pagina = round((list(detalletrasp.values_list('id', flat=True)).index(detalle.id) + 1) / 9.0)
                try:
                    rest = 0
                    if str(pagina).split('.') > 1:
                        rest = (1 if int(str(pagina).split('.')[1]) > 0 else 0)
                    pagina = int(str(pagina).split('.')[0]) + rest
                except:
                    pass
                return JsonResponse({"result": "ok", "encontrado": True, "pagina": pagina, "id": detalle.id})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al buscar los datos"})

        elif action == 'actualizaestado':
            try:
                detalle = DetalleConstatacionFisica.objects.get(pk=int(request.POST['id']))
                detalle.estadoactual_id = request.POST['valestado']
                detalle.save(request)
                return JsonResponse(
                    {"result": "ok", "reload": 'False', 'contador': detalle.codigoconstatacion.contadores_pertenece()})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos"})

        elif action == 'actualizaestadootro':
            try:
                detalle = DetalleConstatacionFisica.objects.get(pk=int(request.POST['id']))
                detalle.estadoactual_id = request.POST['valestado']
                detalle.save(request)
                return JsonResponse(
                    {"result": "ok", "reload": 'False', 'contador': detalle.codigoconstatacion.contadores_otros()})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos"})

        elif action == 'actualizaobservacion':
            try:
                detalle = DetalleConstatacionFisica.objects.get(pk=int(request.POST['id']))
                detalle.observacion = request.POST['valobser']
                detalle.save(request)
                return JsonResponse(
                    {"result": "ok", "reload": 'False', 'contador': detalle.codigoconstatacion.contadores_pertenece()})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos"})

        elif action == 'detallenoidentificado':
            try:
                constatacion = ConstatacionFisica.objects.get(pk=int(request.POST['id']))
                codigobarra = request.POST['codigo']
                if codigobarra:
                    if constatacion.detallenoidentificado_set.filter(codigobarra=request.POST['codigo']).exists():
                        return JsonResponse({"result": "bad", "mensaje": u"Activo ya adicionado"})
                detalle = DetalleNoIdentificado(codigoconstatacion=constatacion,
                                                codigobarra=request.POST['codigo'],
                                                catalogobien_id=int(request.POST['catalogo']),
                                                serie=request.POST['serie'],
                                                descripcion=request.POST['descripcion'],
                                                modelo=request.POST['modelo'],
                                                marca=request.POST['marca'],
                                                estado_id=int(request.POST['codestado']))
                detalle.save(request)
                detallecons = constatacion.detallenoidentificado_set.all().order_by('codigobarra')
                pagina = ((list(detallecons.values_list('id', flat=True)).index(detalle.id)) / 10.0)
                try:
                    pagina = int(pagina) + (1 if str(pagina).split('.')[1] != '00' else 0)
                except:
                    pass
                return JsonResponse({"result": "ok", "mensaje": u"Modificado por otro usuario", "reload": 'True',
                                     'contador': detalle.codigoconstatacion.contadores_ni(), 'pagina': pagina})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos"})

        elif action == 'editcons':
            try:
                constatacion = ConstatacionFisica.objects.get(pk=request.POST['id'])
                f = ConstatacionForm(request.POST)
                if f.is_valid():
                    constatacion.observacion = f.cleaned_data['observacion']
                    constatacion.save(request)
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
            return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'edittraspaso':
            try:
                f = TraspasoActivoForm(request.POST)
                if f.is_valid():
                    traspaso = TraspasoActivo.objects.get(pk=request.POST['id'])
                    traspaso.usuariobienrecibe = f.cleaned_data['usuariobienrecibe']
                    traspaso.ubicacionbienrecibe = f.cleaned_data['ubicacionbienrecibe']
                    traspaso.custodiobienrecibe = f.cleaned_data['custodiobienrecibe']
                    traspaso.tiposolicitud = f.cleaned_data['tiposolicitud']
                    traspaso.oficio = f.cleaned_data['oficio']
                    traspaso.fechaoficio = f.cleaned_data['fechaoficio']
                    traspaso.solicitante = f.cleaned_data['solicitante']
                    traspaso.observacion = f.cleaned_data['observacion']
                    traspaso.save(request)
                    return JsonResponse({"result": "ok", "id": traspaso.id})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'edittraspasocustodio':
            try:
                f = TraspasoActivoForm(request.POST)
                if f.is_valid():
                    traspaso = TraspasoActivo.objects.get(pk=request.POST['id'])
                    traspaso.custodiobienrecibe = f.cleaned_data['custodiobienrecibe']
                    traspaso.tiposolicitud = f.cleaned_data['tiposolicitud']
                    traspaso.oficio = f.cleaned_data['oficio']
                    traspaso.fechaoficio = f.cleaned_data['fechaoficio']
                    traspaso.solicitante = f.cleaned_data['solicitante']
                    traspaso.observacion = f.cleaned_data['observacion']
                    traspaso.save(request)
                    return JsonResponse({"result": "ok", "id": traspaso.id})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'editbaja':
            try:
                f = BajaActivoForm(request.POST)
                if f.is_valid():
                    baja = BajaActivo.objects.get(pk=request.POST['id'])
                    # baja.usuariorecibe = f.cleaned_data['usuariorecibe']
                    baja.oficio = f.cleaned_data['oficio']
                    baja.tipobaja = f.cleaned_data['tipobaja']
                    baja.fechaoficio = f.cleaned_data['fechaoficio']
                    baja.tiposolicitud = f.cleaned_data['tiposolicitud']
                    baja.solicitante = f.cleaned_data['solicitante']
                    if f.cleaned_data['usuariobienentrega']:
                        baja.usuariobienentrega = f.cleaned_data['usuariobienentrega']
                    if f.cleaned_data['custodioentrega']:
                        baja.custodioentrega = f.cleaned_data['custodioentrega']
                    if f.cleaned_data['ubicacionbienentrega']:
                        baja.ubicacionbienentrega = f.cleaned_data['ubicacionbienentrega']

                    baja.observacion = f.cleaned_data['observacion']
                    # baja.experto = f.cleaned_data['experto']
                    # baja.contadorper = f.cleaned_data['contadorper']
                    baja.save(request)
                    return JsonResponse({"result": "ok", "id": baja.id})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'edittipobaja':
            try:
                f = TipoBajaForm(request.POST)
                if f.is_valid():
                    tipobaja = TipoBaja.objects.get(pk=request.POST['id'])
                    tipobaja.nombre = f.cleaned_data['nombre']
                    tipobaja.save(request)
                    log(u'Modifico Tipo Baja: %s' % tipobaja, request, "edit")
                    return JsonResponse({"result": "ok", "id": tipobaja.id})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'editacta':
            try:
                acta = TraspasoActivo.objects.get(pk=request.POST['id'])
                f = ActasForm(request.POST)
                if f.is_valid():
                    acta.observacion = f.cleaned_data['observacion']
                    acta.proveedor = f.cleaned_data['proveedor']
                    acta.custodiobienrecibe = f.cleaned_data['custodio']
                    acta.save(request)
                    for d in acta.detalletraspasoactivo_set.all():
                        activo = ActivoFijo.objects.get(pk=int(d.activo.id))
                        activo.custodio = f.cleaned_data['custodio']
                        activo.tipocomprobante = f.cleaned_data['tipocomprobante']
                        activo.save(request)
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
            return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'importar':
            try:
                form = ImportarArchivoXLSForm(request.POST, request.FILES)
                if form.is_valid():
                    hoy = datetime.now().date()
                    nfile = request.FILES['archivo']
                    nfile._name = generar_nombre("importacion_", nfile._name)
                    archivo = ArchivoActivoFijo(nombre='IMPORTACION DE ACTIVOS',
                                                fecha=datetime.now().date(),
                                                archivo=nfile,
                                                estado=1,
                                                tipobien=form.cleaned_data['tipobien'])
                    archivo.save(request)
                    return JsonResponse({"result": "ok"})
                else:
                    return JsonResponse({"result": "bad", "mensaje": u"El archivo no tiene el formato correcto."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'exportar':
            try:
                f = ExportacionForm(request.POST)
                if f.is_valid():
                    exportacion = ExportacionesActivos(fecha=f.cleaned_data['fecha'],
                                                       clasebien=f.cleaned_data['clasebien'],
                                                       tipobien=f.cleaned_data['tipobien'],
                                                       cuentacontable=f.cleaned_data['cuentacontable'],
                                                       estado=1)
                    exportacion.save(request)
                    cursor = connection.cursor()
                    if exportacion.tipobien:
                        cursor.execute(
                            "INSERT INTO sagest_exportacionesactivos_activos (exportacionesactivos_id, activofijo_id) "
                            "SELECT %s, sagest_activofijo.id from sagest_activofijo LEFT JOIN sagest_catalogobien ON sagest_activofijo.catalogo_id = sagest_catalogobien.id WHERE statusactivo=1 AND sagest_activofijo.status=True  AND clasebien=%s AND sagest_catalogobien.tipobien_id = %s  AND cuentacontable_id=%s AND subidogobierno=False;",
                            [exportacion.id, exportacion.clasebien, exportacion.tipobien.id,
                             exportacion.cuentacontable.id])
                    else:
                        cursor.execute(
                            "INSERT INTO sagest_exportacionesactivos_activos (exportacionesactivos_id, activofijo_id) "
                            "SELECT %s, sagest_activofijo.id from sagest_activofijo LEFT JOIN sagest_catalogobien ON sagest_activofijo.catalogo_id = sagest_catalogobien.id WHERE statusactivo=1 AND sagest_activofijo.status=True  AND clasebien=%s AND cuentacontable_id=%s AND subidogobierno=False;",
                            [exportacion.id, exportacion.clasebien, exportacion.cuentacontable.id])
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
            return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'generarexpo':
            try:
                exportacion = ExportacionesActivos.objects.get(pk=request.POST['id'])
                output_folder = os.path.join(os.path.join(SITE_ROOT, 'media', 'activos'))
                try:
                    os.makedirs(output_folder)
                except Exception as ex:
                    pass
                nombre = 'ACTIVOS_' + elimina_tildes(exportacion.fecha.strftime(
                    '%Y%m%d_%H%M%S') + '_' + exportacion.clase() + '_' + exportacion.cuentacontable.descripcion).replace(
                    ' ', '') + ".csv"
                filename = os.path.join(output_folder, nombre)
                with io.open(filename, 'wt', encoding="ascii", errors="replace") as fichero:
                    linea = 1
                    cantidad_total = exportacion.activos.count()
                    for activo in exportacion.activos.all():
                        if linea == 1:
                            pass
                        # if exportacion.tipobien.id == 1:
                        if exportacion:
                            colornombre = ''
                            if activo.color:
                                colornombre = activo.color.nombre

                            fila = u"2-Bienes Muebles,181,0000,0000,5,BLD,%s,%s,,%s,%s,%s,1,Matriz,15-ACTAS DE ENTREGA-RECEPCION,01-Acta de Entrega-Recepcion,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,,,,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s" % (
                            activo.fechaingreso.strftime("%d/%m/%Y"),
                            activo.activo_origen(),
                            activo.catalogo.descripcion.replace(";", " ").replace(":", " ").replace("'", " ").replace(
                                "-", " ").replace("/", " ").replace("&", " ").replace("+", " ").replace("{",
                                                                                                        " ").replace(
                                "}", " ").replace("*", " ").replace("=", " ").replace('"', ' ').replace(',',
                                                                                                        ' ').encode(
                                "ascii", "ignore"),
                            activo.catalogo.identificador,
                            activo.descripcion.replace(";", " ").replace(":", " ").replace("'", " ").replace("-",
                                                                                                             " ").replace(
                                "/", " ").replace("&", " ").replace("+", " ").replace("{", " ").replace("}",
                                                                                                        " ").replace(
                                "*", " ").replace("=", " ").replace('"', ' ').replace(',', ' ').encode("ascii",
                                                                                                       "ignore"),
                            activo.tipodoc(),
                            activo.fechacomprobante.strftime("%d/%m/%Y"),
                            activo.codigointerno,
                            activo.activo_estado(),
                            str(activo.costo),
                            activo.activo_deprecia(),
                            activo.responsable.identificacion(),
                            activo.responsable.nombre_completo().encode("ascii", "ignore"),
                            activo.ubicacion.codigo,
                            activo.serie.replace("'", "").replace(";", " ").replace(":", " ").replace("'", " ").replace(
                                "-", " ").replace("/", " ").replace("&", " ").replace("+", " ").replace("{",
                                                                                                        " ").replace(
                                "}", " ").replace("*", " ").replace("=", " ").replace('"', ' ').replace(',',
                                                                                                        ' ').encode(
                                "ascii", "ignore"),
                            activo.modelo.replace("'", "").replace(";", " ").replace(":", " ").replace("'",
                                                                                                       " ").replace("-",
                                                                                                                    " ").replace(
                                "/", " ").replace("&", " ").replace("+", " ").replace("{", " ").replace("}",
                                                                                                        " ").replace(
                                "*", " ").replace("=", " ").replace('"', ' ').replace(',', ' ').encode("ascii",
                                                                                                       "ignore"),
                            activo.marca.replace(";", " ").replace(":", " ").replace("'", " ").replace("-",
                                                                                                       " ").replace("/",
                                                                                                                    " ").replace(
                                "&", " ").replace("+", " ").replace("{", " ").replace("}", " ").replace("*",
                                                                                                        " ").replace(
                                "=", " ").replace('"', ' ').replace(',', ' ').encode("ascii", "ignore"),
                            activo.cuentacontable.cuenta,
                            str(activo.costo),
                            str(activo.valorresidual),
                            str(activo.valorlibros),
                            str(activo.valordepreciacionacumulada),
                            (activo.fechafindeprec.strftime("%d/%m/%Y") if activo.deprecia else ""),
                            (str(activo.vidautil) if activo.deprecia else "0"),
                            colornombre.replace(";", " ").replace(":", " ").replace("'", " ").replace("-", " ").replace(
                                "/", " ").replace("&", " ").replace("+", " ").replace("{", " ").replace("}",
                                                                                                        " ").replace(
                                "*", " ").replace("=", " ").replace('"', ' ').replace(',', ' ').encode("ascii",
                                                                                                       "ignore"),
                            activo.material.replace(";", " ").replace(":", " ").replace("'", " ").replace("-",
                                                                                                          " ").replace(
                                "/", " ").replace("&", " ").replace("+", " ").replace("{", " ").replace("}",
                                                                                                        " ").replace(
                                "*", " ").replace("=", " ").replace('"', ' ').replace(',', ' ').encode("ascii",
                                                                                                       "ignore"),
                            activo.dimensiones.replace(";", " ").replace(":", " ").replace("'", " ").replace("-",
                                                                                                             " ").replace(
                                "/", " ").replace("&", " ").replace("+", " ").replace("{", " ").replace("}",
                                                                                                        " ").replace(
                                "*", " ").replace("=", " ").replace('"', ' ').replace(',', ' ').encode("ascii",
                                                                                                       "ignore"),
                            activo.observacion.replace(";", " ").replace(":", " ").replace("'", " ").replace("-",
                                                                                                             " ").replace(
                                "/", " ").replace("&", " ").replace("+", " ").replace("{", " ").replace("}",
                                                                                                        " ").replace(
                                "*", " ").replace("=", " ").replace('"', ' ').replace(',', ' ').encode("ascii",
                                                                                                       "ignore"))
                            fila = fila.replace('\n', ' ').replace('\r', ' ')
                            if not linea == cantidad_total:
                                fila += '\n'
                            fichero.write(fila)
                            linea += 1
                fichero.close()
                exportacion.ficherocsv.name = "activos/%s" % nombre
                exportacion.save(request)
                return JsonResponse({'result': 'ok', 'archivo': nombre})
            except Exception as ex:
                transaction.set_rollback(True)
                translator = Translator()
                return JsonResponse({"result": "bad", "mensaje": translator.translate(ex.__str__(), 'es').text})

        elif action == 'generarexpoxls':
            try:
                exportacion = ExportacionesActivos.objects.get(pk=request.POST['id'])
                output_folder = os.path.join(os.path.join(SITE_ROOT, 'media', 'activos'))
                style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                style_nb = easyxf('font: name Times New Roman, color-index blue, bold on',
                                  num_format_str='#,##0.00')
                style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                title = easyxf(
                    'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                style1 = easyxf(num_format_str='D-MMM-YY')
                font_style = XFStyle()
                font_style.font.bold = True
                font_style2 = XFStyle()
                font_style2.font.bold = False
                wb = Workbook(encoding='utf-8')
                ws = wb.add_sheet('exp_xls_post_part')
                nombre = 'ACTIVOS_' + elimina_tildes(exportacion.fecha.strftime(
                    '%Y%m%d_%H%M%S') + '_' + exportacion.clase() + '_' + exportacion.cuentacontable.descripcion).replace(
                    ' ', '') + ".xls"
                filename = os.path.join(output_folder, nombre)
                book = xlwt.Workbook()
                date_format = xlwt.XFStyle()
                date_format.num_format_str = 'dd/mm/yyyy'
                sheet1 = book.add_sheet('datos')
                fila = 0
                for r in exportacion.activos.all():
                    i = 0
                    ws.write(fila, 0, '2-Bienes Muebles', font_style2)
                    ws.write(fila, 1, '181', font_style2)
                    ws.write(fila, 2, '0000', font_style2)
                    ws.write(fila, 3, '0000', font_style2)
                    ws.write(fila, 4, '5', font_style2)
                    ws.write(fila, 5, 'BLD', font_style2)
                    ws.write(fila, 6, r.fechaingreso, style1)
                    ws.write(fila, 7, r.activo_origen(), font_style2)
                    ws.write(fila, 8, ' ', font_style2)
                    ws.write(fila, 9, r.catalogo.descripcion, font_style2)
                    ws.write(fila, 10, r.catalogo.identificador, font_style2)
                    ws.write(fila, 11, r.descripcion, font_style2)
                    ws.write(fila, 12, '1', font_style2)
                    ws.write(fila, 13, 'Matriz', font_style2)
                    ws.write(fila, 14, '15-ACTAS DE ENTREGA-RECEPCION', font_style2)
                    ws.write(fila, 15, '01-Acta de Entrega-Recepcion', font_style2)
                    ws.write(fila, 16, r.tipodoc(), font_style2)
                    ws.write(fila, 17, r.fechacomprobante, style1)
                    ws.write(fila, 18, r.codigointerno, font_style2)
                    ws.write(fila, 19, r.activo_estado(), font_style2)
                    ws.write(fila, 20, str(r.costo), font_style2)
                    ws.write(fila, 21, r.activo_deprecia(), font_style2)
                    ws.write(fila, 22, r.responsable.identificacion(), font_style2)
                    ws.write(fila, 23, r.responsable.nombre_completo(), font_style2)
                    ws.write(fila, 24, r.ubicacion.codigo, font_style2)
                    ws.write(fila, 25, r.serie, font_style2)
                    ws.write(fila, 26, r.modelo, font_style2)
                    ws.write(fila, 27, r.marca, font_style2)
                    ws.write(fila, 28, ' ', font_style2)
                    ws.write(fila, 29, ' ', font_style2)
                    ws.write(fila, 30, ' ', font_style2)
                    ws.write(fila, 31, r.cuentacontable.cuenta, font_style2)
                    ws.write(fila, 32, str(r.costo), font_style2)
                    ws.write(fila, 33, str(r.valorresidual), font_style2)
                    ws.write(fila, 34, str(r.valorlibros), font_style2)
                    ws.write(fila, 35, str(r.valordepreciacionacumulada), font_style2)
                    ws.write(fila, 36, (r.fechafindeprec if r.deprecia else ""), style1)
                    ws.write(fila, 37, (str(r.vidautil) if r.deprecia else "0"), font_style2)
                    colornombre = ''
                    if r.color:
                        colornombre = r.color.nombre

                    ws.write(fila, 38, colornombre, font_style2)
                    ws.write(fila, 39, r.material, font_style2)
                    ws.write(fila, 40, r.dimensiones, font_style2)
                    ws.write(fila, 41, r.observacion, font_style2)
                    fila += 1
                wb.save(filename)
                exportacion.ficheroxls.name = "activos/%s" % nombre
                exportacion.save(request)
                return JsonResponse({'result': 'ok', 'archivo': nombre})
            except Exception as ex:
                transaction.set_rollback(True)
                translator = Translator()
                return JsonResponse({"result": "bad", "mensaje": translator.translate(ex.__str__(), 'es').text})

        elif action == 'subir':
            try:
                form = ImportarArchivoXLSForm(request.POST, request.FILES)
                if form.is_valid():
                    importacion = ArchivoActivoFijo.objects.get(pk=int(request.POST['id']))
                    hoy = datetime.now().date()
                    nfile = request.FILES['archivo']
                    nfile._name = generar_nombre("importacion_", nfile._name)
                    importacion.archivo = nfile
                    importacion.save(request)
                    log(u'Modifico archivo de importacion: %s' % importacion, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    return JsonResponse({"result": "bad", "mensaje": u"El archivo no tiene el formato correcto."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'informebajapdf':
            mensaje = "Problemas al generar informe de baja."
            try:
                bajaactivo = None
                detalle = None
                director = None

                activofijo = ActivoFijo.objects.get(pk=request.POST['id'])
                if activofijo.informeactivobaja_set.filter(status=True):
                    bajaactivo = activofijo.informeactivobaja_set.filter(status=True)[0]
                    perso = Persona.objects.get(usuario=bajaactivo.usuario_creacion)
                    detalle = bajaactivo.detalleinformeactivobaja_set.filter(status=True)
                    if DirectorResponsableBaja.objects.filter(actual=True, status=True).exists():
                        director = DirectorResponsableBaja.objects.get(actual=True, status=True)
                        if not bajaactivo.fecha_creacion.date() > director.fechainicio:
                            if DirectorResponsableBaja.objects.filter(fechainicio__lte=bajaactivo.fecha_creacion.date(),
                                                                      fechafin__gte=bajaactivo.fecha_creacion.date(),
                                                                      status=True).exists():
                                director = DirectorResponsableBaja.objects.get(
                                    fechainicio__lte=bajaactivo.fecha_creacion.date(),
                                    fechafin__gte=bajaactivo.fecha_creacion.date(), status=True)
                if bajaactivo.tipoinforme == 1:
                    dir = 'inventario_activofijo/informebajapdf.html'
                else:
                    if PersonaDepartamentoFirmas.objects.filter(actualidad=True, status=True).exists():
                        director = PersonaDepartamentoFirmas.objects.get(actualidad=True, status=True,
                                                                         tipopersonadepartamento_id=1,
                                                                         departamentofirma_id=2)

                    if PersonaDepartamentoFirmas.objects.filter(status=True,
                                                                fechafin__gte=bajaactivo.fecha_creacion.date(),
                                                                fechainicio__lte=bajaactivo.fecha_creacion.date(),
                                                                tipopersonadepartamento_id=1,
                                                                departamentofirma_id=2).exists():
                        director = PersonaDepartamentoFirmas.objects.get(status=True,
                                                                         fechafin__gte=bajaactivo.fecha_creacion.date(),
                                                                         fechainicio__lte=bajaactivo.fecha_creacion.date(),
                                                                         tipopersonadepartamento_id=1,
                                                                         departamentofirma_id=2)

                    dir = 'inventario_activofijo/informebajapdfmantenimiento.html'

                return conviert_html_to_pdf(dir,
                                            {'pagesize': 'A4',
                                             'informe': activofijo, 'perso': perso, 'director': director,
                                             'bajaactivo': bajaactivo, 'detalle': detalle, 'hoy': datetime.now().date()
                                             })
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos. %s" % ex})

        elif action == 'procesarimp':
            try:
                importacion = ArchivoActivoFijo.objects.get(pk=request.POST['id'])
                workbook = load_workbook(importacion.archivo.file.name)
                workbookincidencias = xlwt.Workbook()
                sheetincidencias = workbookincidencias.add_sheet('Incidencias')
                sheetcorrectos = workbookincidencias.add_sheet('Activosguardados')
                sheet = workbook._sheets[0]
                listaactas = []
                if importacion.tipobien.id == 2:
                    # VERIFICACIÓN
                    # VEHICULOS
                    linea = 1
                    errorfichero = False
                    for rowx in sheet.rows:
                        errores = False
                        if linea >= 3:
                            codigo = str(sheet.cell(row=linea, column=1).value)
                            cedula = str(sheet.cell(row=linea, column=26).value)
                            tipo = str(sheet.cell(row=linea, column=43).value)
                            if not Persona.objects.filter(cedula__icontains=cedula).exists():
                                sheetincidencias.write(linea, 0, codigo)
                                sheetincidencias.write(linea, 1, 'NO EXISTE EL USUARIO')
                                errores = True
                                errorfichero = True
                            elif not CuentaContable.objects.filter(
                                    cuenta=sheet.cell(row=linea, column=31).value).exists():
                                sheetincidencias.write(linea, 0, codigo)
                                sheetincidencias.write(linea, 1, 'NO EXISTE LA CUENTA CONTABLE')
                                errores = True
                                errorfichero = True
                            elif sheet.cell(row=linea, column=42).value is None:
                                sheetincidencias.write(linea, 0, codigo)
                                sheetincidencias.write(linea, 1, 'NO TIENE DESCRIPCION')
                                errores = True
                                errorfichero = True
                            elif sheet.cell(row=linea, column=46).value is None:
                                sheetincidencias.write(linea, 0, codigo)
                                sheetincidencias.write(linea, 1, 'NO TIENE OBSERVACION')
                                errores = True
                                errorfichero = True
                            elif sheet.cell(row=linea, column=44).value is None:
                                sheetincidencias.write(linea, 0, codigo)
                                sheetincidencias.write(linea, 1, 'NO TIENE N° COMPROBANTE')
                                errores = True
                                errorfichero = True
                            elif sheet.cell(row=linea, column=45).value is None:
                                sheetincidencias.write(linea, 0, codigo)
                                sheetincidencias.write(linea, 1, 'NO TIENE FECHA COMPROBANTE')
                                errores = True
                                errorfichero = True
                            elif not TipoDocumento.objects.filter(codigo__icontains=tipo).exists():
                                sheetincidencias.write(linea, 0, codigo)
                                sheetincidencias.write(linea, 1, 'NO TIENE TIPO COMPROBANTE')
                                errores = True
                                errorfichero = True
                        linea += 1
                    nombre = 'incidencias_activos_' + str(importacion.id) + '.xls'
                    nombrearchivo = os.path.join(MEDIA_ROOT, 'activos', nombre)
                    workbookincidencias.save(nombrearchivo)
                    importacion.archivoincidencias.name = "activos/%s" % nombre
                    importacion.save(request)
                    if errorfichero:
                        return JsonResponse(
                            {"result": "ok", "obs": True, "archivo": importacion.archivoincidencias.url})
                    # ingreso
                    linea = 1
                    contadorguardados = 0
                    for rowx in sheet.rows:
                        if linea >= 3:
                            habilitado = sheet.cell(row=linea, column=47).value
                            numeroacta = sheet.cell(row=linea, column=4).value
                            if habilitado.upper() in ('HABILITADO', 'S', 'SI', 'APROBADO') and numeroacta:
                                codigo = str(sheet.cell(row=linea, column=1).value)
                                tipo = TipoDocumento.objects.filter(
                                    codigo__icontains=str(sheet.cell(row=linea, column=43).value))[0]
                                if not ActivoFijo.objects.filter(codigogobierno=codigo).exists():
                                    codigocatalogo = str(sheet.cell(row=linea, column=3).value)
                                    if not CatalogoBien.objects.filter(identificador=codigocatalogo).exists():
                                        catalogo = CatalogoBien(identificador=codigocatalogo,
                                                                descripcion=str(sheet.cell(row=linea, column=6).value),
                                                                tipobien_id=2)
                                        catalogo.save(request)
                                        log(u'Adicionó catalogo del bien: %s' % catalogo, request, "add")
                                    else:
                                        catalogo = CatalogoBien.objects.filter(identificador=codigocatalogo)[0]
                                    cedula = str(sheet.cell(row=linea, column=26).value)
                                    usuario = Persona.objects.filter(cedula__icontains=cedula)[0]
                                    if not Ubicacion.objects.filter(
                                            nombre=sheet.cell(row=linea, column=25).value).exists():
                                        ubicacion = Ubicacion(nombre=sheet.cell(row=linea, column=25).value,
                                                              codigo='',
                                                              observacion='')
                                        ubicacion.save(request)
                                        log(u'Adicionó ubicación: %s' % ubicacion, request, "add")
                                    else:
                                        ubicacion = \
                                        Ubicacion.objects.filter(nombre=sheet.cell(row=linea, column=25).value)[0]
                                    custodio = ubicacion.responsable
                                    if not ClaseVehiculo.objects.filter(
                                            nombre=sheet.cell(row=linea, column=14).value).exists():
                                        clasevehiculo = ClaseVehiculo(nombre=sheet.cell(row=linea, column=14).value)
                                        clasevehiculo.save(request)
                                        log(u'Adicionó clase de vehículo: %s' % clasevehiculo, request, "add")
                                    else:
                                        clasevehiculo = \
                                        ClaseVehiculo.objects.filter(nombre=sheet.cell(row=linea, column=14).value)[0]
                                    if not TipoVehiculo.objects.filter(
                                            nombre__icontains=sheet.cell(row=linea, column=15).value).exists():
                                        tipovehiculo = TipoVehiculo(nombre=sheet.cell(row=linea, column=15).value)
                                        tipovehiculo.save(request)
                                        log(u'Adicionó tipo de vehículo: %s' % tipovehiculo, request, "add")
                                    else:
                                        tipovehiculo = \
                                        TipoVehiculo.objects.filter(nombre=sheet.cell(row=linea, column=15).value)[0]
                                    if not EstadoProducto.objects.filter(
                                            nombre=sheet.cell(row=linea, column=22).value).exists():
                                        estado = EstadoProducto(nombre=sheet.cell(row=linea, column=22).value)
                                        estado.save(request)
                                        log(u'Adicionó tipo de vehículo: %s' % tipovehiculo, request, "add")
                                    else:
                                        estado = \
                                        EstadoProducto.objects.filter(nombre=sheet.cell(row=linea, column=22).value)[0]
                                    if not Color.objects.filter(nombre=sheet.cell(row=linea, column=20).value).exists():
                                        colorprimario = Color(nombre=sheet.cell(row=linea, column=20).value)
                                        colorprimario.save(request)
                                        log(u'Adicionó color: %s' % colorprimario, request, "add")
                                    else:
                                        colorprimario = \
                                        Color.objects.filter(nombre=sheet.cell(row=linea, column=20).value)[0]
                                    if not Color.objects.filter(nombre=sheet.cell(row=linea, column=21).value).exists():
                                        colorsecundario = Color(nombre=sheet.cell(row=linea, column=21).value)
                                        colorsecundario.save(request)
                                        log(u'Adicionó color: %s' % colorprimario, request, "add")
                                    else:
                                        colorsecundario = \
                                        Color.objects.filter(nombre=sheet.cell(row=linea, column=21).value)[0]
                                    if not OrigenIngreso.objects.filter(
                                            nombre=sheet.cell(row=linea, column=28).value).exists():
                                        origeningreso = OrigenIngreso(nombre=sheet.cell(row=linea, column=28).value)
                                        origeningreso.save(request)
                                        log(u'Adicionó color: %s' % origeningreso, request, "add")
                                    else:
                                        origeningreso = \
                                        OrigenIngreso.objects.filter(nombre=sheet.cell(row=linea, column=28).value)[0]
                                    if not sheet.cell(row=linea, column=45).value is None:
                                        fechacomp = sheet.cell(row=linea, column=45).value.date()
                                    else:
                                        fechacomp = None
                                    cuenta = \
                                    CuentaContable.objects.filter(cuenta=sheet.cell(row=linea, column=31).value)[0]
                                    if TraspasoActivo.objects.filter(tipo=1, usuariobienrecibe=usuario,
                                                                     ubicacionbienrecibe=ubicacion, estado=1,
                                                                     archivoactivofijo=importacion).exists():
                                        acta = TraspasoActivo.objects.filter(tipo=1, usuariobienrecibe=usuario,
                                                                             ubicacionbienrecibe=ubicacion, estado=1,
                                                                             archivoactivofijo=importacion)[0]
                                    else:
                                        acta = TraspasoActivo(tipo=1,
                                                              actaentregagobierno=int(numeroacta),
                                                              fecha=datetime.now(),
                                                              usuariobienrecibe=usuario,
                                                              ubicacionbienrecibe=ubicacion,
                                                              responsablebienes_id=RESPONSABLE_BIENES_ID,
                                                              asistentebodega_id=ASISTENTE_BODEGA_ID,
                                                              estado=1,
                                                              custodiobienrecibe=custodio)
                                        acta.save(request)
                                        log(u'Adicionó un traspaso de activo: %s' % acta, request, "add")
                                        secuencia = secuencia_activos(request)
                                        secuencia.numeroasignacion += 1
                                        secuencia.save(request)
                                        acta.numero = secuencia.numeroasignacion
                                        acta.normativaacta = secuencia.normativaacta
                                        acta.save(request)
                                    activo = ActivoFijo(codigogobierno=codigo,
                                                        catalogo=catalogo,
                                                        fechaingreso=sheet.cell(row=linea, column=33).value.date(),
                                                        serie=sheet.cell(row=linea, column=7).value,
                                                        modelo=sheet.cell(row=linea, column=8).value,
                                                        marca=sheet.cell(row=linea, column=9).value,
                                                        clasevehiculo=clasevehiculo,
                                                        tipovehiculo=tipovehiculo,
                                                        custodio=custodio,
                                                        responsable=usuario,
                                                        descripcion=str(sheet.cell(row=linea, column=42).value),
                                                        observacion=str(sheet.cell(row=linea, column=46).value),
                                                        tipocomprobante=tipo,
                                                        numerocomprobante=str(sheet.cell(row=linea, column=44).value),
                                                        fechacomprobante=fechacomp,
                                                        ubicacion=ubicacion,
                                                        numeromotor=str(sheet.cell(row=linea, column=16).value),
                                                        numerochasis=str(sheet.cell(row=linea, column=17).value),
                                                        aniofabricacion=int(sheet.cell(row=linea, column=18).value),
                                                        placa=str(sheet.cell(row=linea, column=19).value),
                                                        colorprimario=colorprimario,
                                                        colorsecundario=colorsecundario,
                                                        estado=estado,
                                                        origeningreso=origeningreso,
                                                        fechaultimadeprec=sheet.cell(row=linea, column=34).value.date(),
                                                        vidautil=int(sheet.cell(row=linea, column=35).value),
                                                        cuentacontable=cuenta,
                                                        costo=float(sheet.cell(row=linea, column=37).value),
                                                        valorresidual=float(sheet.cell(row=linea, column=38).value),
                                                        valorlibros=float(sheet.cell(row=linea, column=39).value),
                                                        statusactivo=1,
                                                        valordepreciacionacumulada=float(
                                                            sheet.cell(row=linea, column=40).value))
                                    activo.save(request)
                                    log(u'Adicionó activo: %s' % activo, request, "add")
                                    detalle = DetalleTraspasoActivo(codigotraspaso=acta, activo=activo)
                                    detalle.save(request)
                                    importacion.actas.add(acta)
                                    contadorguardados += 1
                                    for acta in importacion.actas.all():
                                        listaactas.append(acta.numero)
                        linea += 1
                elif importacion.tipobien.id == 6:
                    # libros
                    # verificación
                    linea = 1
                    errorfichero = False
                    for rowx in sheet.rows:
                        errores = False
                        if linea >= 3:
                            codigo = str(sheet.cell(row=linea, column=1).value)
                            cedula = str(sheet.cell(row=linea, column=24).value)
                            tipo = str(sheet.cell(row=linea, column=42).value)
                            if not Persona.objects.filter(cedula__icontains=cedula).exists():
                                sheetincidencias.write(linea, 0, codigo)
                                sheetincidencias.write(linea, 1, 'NO EXISTE EL USUARIO')
                                errores = True
                                errorfichero = True
                            elif not CuentaContable.objects.filter(
                                    cuenta=sheet.cell(row=linea, column=29).value).exists():
                                sheetincidencias.write(linea, 0, codigo)
                                sheetincidencias.write(linea, 1, 'NO EXISTE LA CUENTA CONTABLE')
                                errores = True
                                errorfichero = True
                            elif len(remover_caracteres_especiales_unicode(
                                    sheet.cell(row=linea, column=40).value).strip()) == 0:
                                sheetincidencias.write(linea, 0, codigo)
                                sheetincidencias.write(linea, 1, 'NO TIENE DESCRIPCION')
                                errores = True
                                errorfichero = True
                            elif len(remover_caracteres_especiales_unicode(
                                    sheet.cell(row=linea, column=41).value).strip()) == 0:
                                sheetincidencias.write(linea, 0, codigo)
                                sheetincidencias.write(linea, 1, 'NO TIENE OBSERVACION')
                                errores = True
                                errorfichero = True
                            elif len(str(sheet.cell(row=linea, column=43).value).strip()) == 0:
                                sheetincidencias.write(linea, 0, codigo)
                                sheetincidencias.write(linea, 1, 'NO TIENE N° COMPROBANTE')
                                errores = True
                                errorfichero = True
                            elif not Ubicacion.objects.filter(nombre=sheet.cell(row=linea, column=23).value,
                                                              responsable__isnull=False).exists():
                                sheetincidencias.write(linea, 0, codigo)
                                sheetincidencias.write(linea, 1, sheet.cell(row=linea, column=6).value)
                                sheetincidencias.write(linea, 2,
                                                       'CODIGO BIEN SIN UBICACION O RESPONSABLE, VERIFICAR QUE EXISTA LA UBIACION CON RESPONSABLE')
                                errores = True
                                errorfichero = True
                            elif len(str(sheet.cell(row=linea, column=44).value).strip()) == 0:
                                sheetincidencias.write(linea, 0, codigo)
                                sheetincidencias.write(linea, 1, 'NO TIENE FECHA COMPROBANTE')
                                errores = True
                                errorfichero = True
                            elif not TipoDocumento.objects.filter(codigo__icontains=tipo).exists():
                                sheetincidencias.write(linea, 0, codigo)
                                sheetincidencias.write(linea, 1, 'NO TIENE TIPO COMPROBANTE')
                                errores = True
                                errorfichero = True
                        linea += 1
                    nombre = 'incidencias_activos_' + str(importacion.id) + '.xls'
                    nombrearchivo = os.path.join(MEDIA_ROOT, 'activos', nombre)
                    workbookincidencias.save(nombrearchivo)
                    importacion.archivoincidencias.name = "activos/%s" % nombre
                    importacion.save(request)
                    if errorfichero:
                        return JsonResponse(
                            {"result": "ok", "obs": True, "archivo": importacion.archivoincidencias.url})
                    # ingreso
                    linea = 1
                    contadorguardados = 0
                    for rowx in sheet.rows:
                        # cols = sheet.row_values(rowx)
                        if linea >= 3:
                            habilitado = sheet.cell(row=linea, column=21).value
                            numeroacta = sheet.cell(row=linea, column=4).value

                            if habilitado.upper() in ('HABILITADO', 'S', 'SI', 'APROBADO') and numeroacta:
                                codigo = str(sheet.cell(row=linea, column=1).value)
                                tipo = TipoDocumento.objects.filter(
                                    codigo__icontains=str(sheet.cell(row=linea, column=42).value))[0]
                                if not ActivoFijo.objects.filter(codigogobierno=codigo).exists():
                                    codigocatalogo = str(sheet.cell(row=linea, column=3).value)
                                    if not CatalogoBien.objects.filter(identificador=codigocatalogo).exists():
                                        catalogo = CatalogoBien(identificador=codigocatalogo,
                                                                descripcion=str(sheet.cell(row=linea, column=6).value),
                                                                tipobien_id=6)
                                        catalogo.save(request)
                                    else:
                                        catalogo = CatalogoBien.objects.filter(identificador=codigocatalogo)[0]
                                    cedula = str(sheet.cell(row=linea, column=24).value)
                                    usuario = Persona.objects.filter(cedula__icontains=cedula)[0]
                                    if not Ubicacion.objects.filter(
                                            nombre=sheet.cell(row=linea, column=23).value).exists():
                                        ubicacion = Ubicacion(nombre=sheet.cell(row=linea, column=23).value,
                                                              codigo='',
                                                              observacion='')
                                        ubicacion.save(request)
                                    else:
                                        ubicacion = \
                                        Ubicacion.objects.filter(nombre=sheet.cell(row=linea, column=23).value)[0]
                                    custodio = ubicacion.responsable
                                    if not EstadoProducto.objects.filter(
                                            nombre=sheet.cell(row=linea, column=20).value).exists():
                                        estado = EstadoProducto(nombre=sheet.cell(row=linea, column=20).value)
                                        estado.save(request)
                                    else:
                                        estado = \
                                        EstadoProducto.objects.filter(nombre=sheet.cell(row=linea, column=20).value)[0]
                                    if not OrigenIngreso.objects.filter(
                                            nombre=sheet.cell(row=linea, column=26).value).exists():
                                        origeningreso = OrigenIngreso(nombre=sheet.cell(row=linea, column=26).value)
                                        origeningreso.save(request)
                                    else:
                                        origeningreso = \
                                        OrigenIngreso.objects.filter(nombre=sheet.cell(row=linea, column=26).value)[0]
                                    if not sheet.cell(row=linea, column=44).value is None:
                                        fechacomp = sheet.cell(row=linea, column=44).value.date()
                                    else:
                                        fechacomp = None
                                    cuenta = \
                                    CuentaContable.objects.filter(cuenta=sheet.cell(row=linea, column=29).value)[0]
                                    ultimadeprec = None
                                    if sheet.cell(row=linea, column=32).value:
                                        ultimadeprec = sheet.cell(row=linea, column=32).value.date()
                                    if TraspasoActivo.objects.filter(tipo=1, usuariobienrecibe=usuario,
                                                                     ubicacionbienrecibe=ubicacion, estado=1,
                                                                     archivoactivofijo=importacion).exists():
                                        acta = TraspasoActivo.objects.filter(tipo=1, usuariobienrecibe=usuario,
                                                                             ubicacionbienrecibe=ubicacion, estado=1,
                                                                             archivoactivofijo=importacion)[0]
                                    else:
                                        acta = TraspasoActivo(tipo=1,
                                                              actaentregagobierno=int(numeroacta),
                                                              fecha=datetime.now(),
                                                              usuariobienrecibe=usuario,
                                                              ubicacionbienrecibe=ubicacion,
                                                              responsablebienes_id=RESPONSABLE_BIENES_ID,
                                                              asistentebodega_id=ASISTENTE_BODEGA_ID,
                                                              estado=1,
                                                              custodiobienrecibe=custodio)
                                        acta.save(request)
                                        secuencia = secuencia_activos(request)
                                        secuencia.numeroasignacion += 1
                                        secuencia.save(request)
                                        acta.numero = secuencia.numeroasignacion
                                        acta.normativaacta = secuencia.normativaacta
                                        acta.save(request)

                                    activo = ActivoFijo(codigogobierno=codigo,
                                                        catalogo=catalogo,
                                                        fechaingreso=sheet.cell(row=linea, column=31).value.date(),
                                                        custodio=custodio,
                                                        responsable=usuario,
                                                        ubicacion=ubicacion,
                                                        tipocomprobante=tipo,
                                                        descripcion=remover_caracteres_especiales_unicode(
                                                            sheet.cell(row=linea, column=40).value),
                                                        observacion=remover_caracteres_especiales_unicode(
                                                            sheet.cell(row=linea, column=41).value),
                                                        numerocomprobante=str(sheet.cell(row=linea, column=43).value),
                                                        fechacomprobante=fechacomp,
                                                        serie=str(sheet.cell(row=linea, column=7).value),
                                                        modelo=str(sheet.cell(row=linea, column=8).value),
                                                        marca=str(sheet.cell(row=linea, column=9).value),
                                                        titulo=remover_caracteres_especiales_unicode(
                                                            sheet.cell(row=linea, column=14).value),
                                                        autor=sheet.cell(row=linea, column=15).value,
                                                        editorial=sheet.cell(row=linea, column=16).value,
                                                        fechaedicion=sheet.cell(row=linea, column=17).value.date(),
                                                        numeroedicion=str(sheet.cell(row=linea, column=18).value),
                                                        clasificacionbibliografica=str(
                                                            sheet.cell(row=linea, column=19).value),
                                                        estado=estado,
                                                        origeningreso=origeningreso,
                                                        fechaultimadeprec=ultimadeprec,
                                                        vidautil=int(sheet.cell(row=linea, column=33).value),
                                                        cuentacontable=cuenta,
                                                        costo=float(sheet.cell(row=linea, column=35).value),
                                                        valorresidual=float(sheet.cell(row=linea, column=36).value),
                                                        valorlibros=float(sheet.cell(row=linea, column=37).value),
                                                        statusactivo=1,
                                                        valordepreciacionacumulada=float(
                                                            sheet.cell(row=linea, column=38).value),
                                                        deprecia=False)
                                    activo.save(request)
                                    detalle = DetalleTraspasoActivo(codigotraspaso=acta,
                                                                    activo=activo)
                                    detalle.save(request)
                                    importacion.actas.add(acta)
                                    contadorguardados += 1
                                    for acta in importacion.actas.all():
                                        listaactas.append(acta.numero)
                        linea += 1
                else:
                    # otros
                    # verificación
                    linea = 1
                    errorfichero = False
                    for rowx in sheet.rows:
                        errores = False
                        if linea >= 3:
                            codigo = str(sheet.cell(row=linea, column=1).value)
                            cedula = str(sheet.cell(row=linea, column=21).value)
                            tipo = str(sheet.cell(row=linea, column=38).value)
                            if not Persona.objects.filter(cedula__icontains=cedula).exists():
                                sheetincidencias.write(linea, 0, codigo)
                                sheetincidencias.write(linea, 1, sheet.cell(row=linea, column=6).value)
                                sheetincidencias.write(linea, 2, 'NO EXISTE EL USUARIO')
                                errores = True
                                errorfichero = True
                            elif not CuentaContable.objects.filter(
                                    cuenta=sheet.cell(row=linea, column=26).value).exists():
                                sheetincidencias.write(linea, 0, codigo)
                                sheetincidencias.write(linea, 1, sheet.cell(row=linea, column=6).value)
                                sheetincidencias.write(linea, 2, 'NO EXISTE LA CUENTA CONTABLE')
                                errores = True
                                errorfichero = True
                            elif len((sheet.cell(row=linea, column=37).value).strip()) == 0:
                                sheetincidencias.write(linea, 0, codigo)
                                sheetincidencias.write(linea, 1, sheet.cell(row=linea, column=6).value)
                                sheetincidencias.write(linea, 2, 'NO TIENE DESCRIPCION')
                                errores = True
                                errorfichero = True
                            elif len((sheet.cell(row=linea, column=41).value).strip()) == 0:
                                sheetincidencias.write(linea, 0, codigo)
                                sheetincidencias.write(linea, 1, sheet.cell(row=linea, column=6).value)
                                sheetincidencias.write(linea, 2, 'NO TIENE OBSERVACION')
                                errores = True
                                errorfichero = True
                            elif len(str(sheet.cell(row=linea, column=39).value).strip()) == 0:
                                sheetincidencias.write(linea, 0, codigo)
                                sheetincidencias.write(linea, 1, sheet.cell(row=linea, column=6).value)
                                sheetincidencias.write(linea, 2, 'NO TIENE N° COMPROBANTE')
                                errores = True
                                errorfichero = True
                            elif len(str(sheet.cell(row=linea, column=40).value).strip()) == 0:
                                sheetincidencias.write(linea, 0, codigo)
                                sheetincidencias.write(linea, 1, sheet.cell(row=linea, column=6).value)
                                sheetincidencias.write(linea, 2, 'NO TIENE FECHA COMPROBANTE')
                                errores = True
                                errorfichero = True
                            elif not Ubicacion.objects.filter(nombre=sheet.cell(row=linea, column=20).value,
                                                              responsable__isnull=False).exists():
                                sheetincidencias.write(linea, 0, codigo)
                                sheetincidencias.write(linea, 1, sheet.cell(row=linea, column=6).value)
                                sheetincidencias.write(linea, 2,
                                                       'CODIGO BIEN SIN UBICACION O RESPONSABLE, VERIFICAR QUE EXISTA LA UBIACION CON RESPONSABLE')
                                errores = True
                                errorfichero = True
                            elif not TipoDocumento.objects.filter(codigo__icontains=tipo).exists():
                                sheetincidencias.write(linea, 0, codigo)
                                sheetincidencias.write(linea, 1, sheet.cell(row=linea, column=6).value)
                                sheetincidencias.write(linea, 2, 'NO TIENE TIPO COMPROBANTE')
                                errores = True
                                errorfichero = True
                        linea += 1
                    nombre = 'incidencias_activos_' + str(importacion.id) + '.xls'
                    nombrearchivo = os.path.join(MEDIA_ROOT, 'activos', nombre)
                    workbookincidencias.save(nombrearchivo)
                    importacion.archivoincidencias.name = "activos/%s" % nombre
                    importacion.save(request)
                    if errorfichero:
                        return JsonResponse(
                            {"result": "ok", "obs": True, "archivo": importacion.archivoincidencias.url})
                    # ingreso
                    linea = 1
                    contadorguardados = 0
                    for rowx in sheet.rows:
                        if linea >= 3:
                            habilitado = sheet.cell(row=linea, column=18).value
                            numeroacta = sheet.cell(row=linea, column=4).value

                            if habilitado.upper() in ('HABILITADO', 'S', 'SI', 'APROBADO') and numeroacta:
                                codigo = str(sheet.cell(row=linea, column=1).value)
                                tipo = TipoDocumento.objects.filter(
                                    codigo__icontains=str(sheet.cell(row=linea, column=38).value))[0]
                                # if not ActivoFijo.objects.filter(codigogobierno=codigo).exists():
                                if Ubicacion.objects.filter(nombre=sheet.cell(row=linea, column=20).value,
                                                            responsable__isnull=False).exists():
                                    codigocatalogo = str(sheet.cell(row=linea, column=3).value)
                                    if not CatalogoBien.objects.filter(identificador=codigocatalogo).exists():
                                        catalogo = CatalogoBien(identificador=codigocatalogo,
                                                                descripcion=str(sheet.cell(row=linea, column=6).value),
                                                                tipobien=importacion.tipobien)
                                        catalogo.save(request)
                                    else:
                                        catalogo = CatalogoBien.objects.filter(identificador=codigocatalogo)[0]
                                    cedula = str(sheet.cell(row=linea, column=21).value)
                                    usuario = Persona.objects.filter(cedula__icontains=cedula)[0]
                                    # if not Ubicacion.objects.filter(nombre=sheet.cell(row=linea, column=20).value).exists():
                                    #     ubicacion = Ubicacion(nombre=sheet.cell(row=linea, column=20).value,
                                    #                           codigo='',
                                    #                           observacion='')
                                    #     ubicacion.save(request)
                                    #     log(u'Adicionó ubicacion: %s' % ubicacion, request, "add")
                                    # else:
                                    #     ubicacion = Ubicacion.objects.filter(nombre=sheet.cell(row=linea, column=20).value)[0]
                                    # custodio = ubicacion.responsable

                                    ubicacion = Ubicacion.objects.filter(nombre=sheet.cell(row=linea, column=20).value)[
                                        0]
                                    custodio = ubicacion.responsable
                                    if not EstadoProducto.objects.filter(
                                            nombre=sheet.cell(row=linea, column=17).value).exists():
                                        estado = EstadoProducto(nombre=sheet.cell(row=linea, column=17).value)
                                        estado.save(request)
                                        log(u'Adicionó estado de producto: %s' % estado, request, "add")
                                    else:
                                        estado = \
                                        EstadoProducto.objects.filter(nombre=sheet.cell(row=linea, column=17).value)[0]
                                    if not OrigenIngreso.objects.filter(
                                            nombre=sheet.cell(row=linea, column=23).value).exists():
                                        origeningreso = OrigenIngreso(nombre=sheet.cell(row=linea, column=23).value)
                                        origeningreso.save(request)
                                        log(u'Adicionó origen de ingreso: %s' % origeningreso, request, "add")
                                    else:
                                        origeningreso = \
                                        OrigenIngreso.objects.filter(nombre=sheet.cell(row=linea, column=23).value)[0]
                                    if not Color.objects.filter(nombre=sheet.cell(row=linea, column=14).value).exists():
                                        color = Color(nombre=sheet.cell(row=linea, column=14).value)
                                        color.save(request)
                                        log(u'Adicionó color de activo: %s' % color, request, "add")
                                    else:
                                        color = Color.objects.filter(nombre=sheet.cell(row=linea, column=14).value)[0]
                                    if not sheet.cell(row=linea, column=40).value is None:
                                        fechacomp = sheet.cell(row=linea, column=40).value.date()
                                    else:
                                        fechacomp = None
                                    cuenta = \
                                    CuentaContable.objects.filter(cuenta=sheet.cell(row=linea, column=26).value)[0]
                                    ultimadeprec = None
                                    if sheet.cell(row=linea, column=29).value:
                                        ultimadeprec = sheet.cell(row=linea, column=29).value.date()
                                    if not ActivoFijo.objects.filter(codigogobierno=codigo).exists():
                                        if TraspasoActivo.objects.filter(tipo=1, usuariobienrecibe=usuario,
                                                                         ubicacionbienrecibe=ubicacion, estado=1,
                                                                         archivoactivofijo=importacion).exists():
                                            acta = TraspasoActivo.objects.filter(tipo=1, usuariobienrecibe=usuario,
                                                                                 ubicacionbienrecibe=ubicacion,
                                                                                 estado=1,
                                                                                 archivoactivofijo=importacion)[0]
                                        else:
                                            acta = TraspasoActivo(tipo=1,
                                                                  actaentregagobierno=int(numeroacta),
                                                                  responsablebienes_id=RESPONSABLE_BIENES_ID,
                                                                  asistentebodega_id=ASISTENTE_BODEGA_ID,
                                                                  fecha=datetime.now(),
                                                                  usuariobienrecibe=usuario,
                                                                  ubicacionbienrecibe=ubicacion,
                                                                  estado=1,
                                                                  custodiobienrecibe=custodio)
                                            acta.save(request)
                                            log(u'Adicionó acta de traspaso de activo: %s' % acta, request, "add")
                                            secuencia = secuencia_activos(request)
                                            secuencia.numeroasignacion += 1
                                            secuencia.save(request)
                                            acta.numero = secuencia.numeroasignacion
                                            acta.normativaacta = secuencia.normativaacta
                                            acta.save(request)
                                        dimensiones = 'SIN ESPECIFICAR'
                                        if not str(sheet.cell(row=linea, column=16).value) == 'None':
                                            dimensiones = sheet.cell(row=linea, column=16).value

                                        activo = ActivoFijo(codigogobierno=codigo,
                                                            catalogo=catalogo,
                                                            fechaingreso=sheet.cell(row=linea, column=28).value.date(),
                                                            custodio=custodio,
                                                            responsable=usuario,
                                                            ubicacion=ubicacion,
                                                            descripcion=(sheet.cell(row=linea, column=37).value),
                                                            observacion=(sheet.cell(row=linea, column=41).value),
                                                            tipocomprobante=tipo,
                                                            numerocomprobante=str(
                                                                sheet.cell(row=linea, column=39).value),
                                                            fechacomprobante=fechacomp,
                                                            serie=str(sheet.cell(row=linea, column=7).value),
                                                            modelo=sheet.cell(row=linea, column=8).value,
                                                            marca=sheet.cell(row=linea, column=9).value,
                                                            color=color,
                                                            material=sheet.cell(row=linea, column=15).value,
                                                            dimensiones=dimensiones,
                                                            estado=estado,
                                                            origeningreso=origeningreso,
                                                            fechaultimadeprec=ultimadeprec,
                                                            vidautil=int(sheet.cell(row=linea, column=30).value),
                                                            cuentacontable=cuenta,

                                                            costo=float(
                                                                (sheet.cell(row=linea, column=32).value).replace(',',
                                                                                                                 '')) if type(
                                                                sheet.cell(row=linea,
                                                                           column=32).value) == 'str' else sheet.cell(
                                                                row=linea, column=32).value,
                                                            valorresidual=float(
                                                                (sheet.cell(row=linea, column=33).value).replace(',',
                                                                                                                 '')) if type(
                                                                sheet.cell(row=linea,
                                                                           column=33).value) == 'str' else sheet.cell(
                                                                row=linea, column=33).value,
                                                            valorlibros=float(
                                                                (sheet.cell(row=linea, column=34).value).replace(',',
                                                                                                                 '')) if type(
                                                                sheet.cell(row=linea,
                                                                           column=34).value) == 'str' else sheet.cell(
                                                                row=linea, column=34).value,

                                                            statusactivo=1,

                                                            valordepreciacionacumulada=float(
                                                                (sheet.cell(row=linea, column=35).value).replace(',',
                                                                                                                 '')) if type(
                                                                sheet.cell(row=linea,
                                                                           column=35).value) == 'str' else float(
                                                                sheet.cell(row=linea, column=35).value)
                                                            )

                                        activo.save(request)
                                        log(u'Adicionó activo: %s' % activo, request, "add")
                                        detalle = DetalleTraspasoActivo(codigotraspaso=acta,
                                                                        activo=activo)
                                        detalle.save(request)
                                        importacion.actas.add(acta)
                                        contadorguardados += 1
                                        for acta in importacion.actas.all():
                                            listaactas.append(acta.numero)
                                    else:
                                        act = ActivoFijo.objects.get(codigogobierno=codigo, status=True)
                                        act.responsable = usuario
                                        act.custodio = custodio
                                        act.ubicacion = ubicacion
                                        act.save(request)
                                        log(u'Actualizó activo: %s' % act, request, "add")

                        linea += 1
                importacion.estado = 2
                importacion.numimportados = contadorguardados
                importacion.save(request)
                archivoactivo = ArchivoActivoFijo.objects.get(pk=request.POST['id'])
                miarchivoupdate = openpyxl.load_workbook(archivoactivo.archivo.file.name)
                sheetes = miarchivoupdate.active
                maximo_column = (sheet.max_column) + 1
                columnfila = maximo_column - 1
                sheetes.cell(row=1, column=maximo_column).value = maximo_column
                sheetes.cell(row=2, column=maximo_column).value = 'accion'

                columcustodio = maximo_column + 1
                columnfilacustodio = columcustodio - 1
                sheetes.cell(row=1, column=columcustodio).value = columcustodio
                sheetes.cell(row=2, column=columcustodio).value = 'custodio'
                miarchivoupdate.save(archivoactivo.archivo.file.name)

                miarchivo = openpyxl.load_workbook(archivoactivo.archivo.file.name)
                hojas = miarchivo.get_sheet_names()
                listadoar = miarchivo.get_sheet_by_name(str(hojas[0]))
                totallista = list(listadoar.rows)
                a = 0
                for filas in totallista[:]:
                    filas
                    a += 1
                    if a > 2:
                        if importacion.tipobien.id == 6:
                            if ActivoFijo.objects.filter(codigogobierno=filas[0].value, status=True).exists():
                                filas[columnfila].value = 'CORRECTO'
                                ubicacion = Ubicacion.objects.filter(nombre=sheet.cell(row=a, column=23).value)[0]
                                custodio = ubicacion.responsable
                                filas[
                                    columnfilacustodio].value = custodio.apellido1 + ' ' + custodio.apellido2 + ' ' + custodio.nombres
                            else:
                                filas[columnfila].value = 'ERROR AL SUBIR'
                                ubicacion = Ubicacion.objects.filter(nombre=sheet.cell(row=a, column=23).value)[0]
                                custodio = ubicacion.responsable
                                filas[
                                    columnfilacustodio].value = custodio.apellido1 + ' ' + custodio.apellido2 + ' ' + custodio.nombres
                        elif importacion.tipobien.id == 2:
                            if ActivoFijo.objects.filter(codigogobierno=filas[0].value, status=True).exists():
                                filas[columnfila].value = 'CORRECTO'
                                ubicacion = Ubicacion.objects.filter(nombre=sheet.cell(row=a, column=25).value)[0]
                                custodio = ubicacion.responsable
                                filas[
                                    columnfilacustodio].value = custodio.apellido1 + ' ' + custodio.apellido2 + ' ' + custodio.nombres
                            else:
                                filas[columnfila].value = 'ERROR AL SUBIR'
                                ubicacion = Ubicacion.objects.filter(nombre=sheet.cell(row=a, column=25).value)[0]
                                custodio = ubicacion.responsable
                                filas[
                                    columnfilacustodio].value = custodio.apellido1 + ' ' + custodio.apellido2 + ' ' + custodio.nombres
                        else:
                            if ActivoFijo.objects.filter(codigogobierno=filas[0].value, status=True).exists():
                                filas[columnfila].value = 'CORRECTO'
                                ubicacion = Ubicacion.objects.filter(nombre=sheet.cell(row=a, column=20).value)[0]
                                custodio = ubicacion.responsable
                                filas[
                                    columnfilacustodio].value = custodio.apellido1 + ' ' + custodio.apellido2 + ' ' + custodio.nombres
                            else:
                                filas[columnfila].value = 'ERROR AL SUBIR'
                                ubicacion = Ubicacion.objects.filter(nombre=sheet.cell(row=a, column=20).value)[0]
                                custodio = ubicacion.responsable
                                filas[
                                    columnfilacustodio].value = custodio.apellido1 + ' ' + custodio.apellido2 + ' ' + custodio.nombres

                miarchivo.save(archivoactivo.archivo.file.name)

                log(u'Proceso archivo de importacion: %s' % importacion, request, "edit")
                return JsonResponse({"result": "ok", "obs": False, "actas": listaactas})
            except Exception as ex:
                transaction.set_rollback(True)
                print(ex)
                print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'confirmacion':
            try:
                traslado = TrasladoMantenimiento.objects.get(pk=int(request.POST['id']))
                traslado.tiporegistro = 2
                secuencia = secuencia_activos(request)
                secuencia.numeromantenimientodefinitivo += 1
                secuencia.save(request)
                traslado.numerodefinitivo = secuencia.numeromantenimientodefinitivo
                traslado.fechadefinitivo = datetime.now().date()
                traslado.save(request)
                log(u'Confirmar planificación: %s' % traslado, request, "edit")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'depreciar':
            try:
                fecha = convertir_fecha(request.POST['fecha'])
                if ActivoFijo.objects.filter(fechaultimadeprec__gt=fecha, statusactivo=1, deprecia=True).exists():
                    activo = ActivoFijo.objects.filter(fechaultimadeprec__gt=fecha, statusactivo=1, deprecia=True)[0]
                    return JsonResponse({"result": "bad",
                                         "mensaje": u"Error la fecha de depreciación es incorrecta: La última depreciación tiene fecha %s" % activo.fechaultimadeprec.strftime(
                                             '%d-%m-%Y')})
                if fecha > datetime.now().date():
                    return JsonResponse({"result": "bad", "mensaje": u"La fecha establecida es mayor a la actual"})
                activo = ActivoFijo.objects.filter(statusactivo=1, deprecia=True).exclude(vidautil=0)
                activos = []
                for ac in activo:
                    act = {'descripcion': unicode(ac.descripcion if len(ac.descripcion) > 0 else "None"), 'id': ac.id}
                    activos.append(act)
                return JsonResponse({"result": "ok", "cantidad": len(activos), "activos": activos})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al generar lista para depreciar"})

        elif action == 'depreciando':
            try:
                fecha = convertir_fecha(request.POST['fecha'])
                activo = ActivoFijo.objects.get(pk=request.POST['maid'])
                activo.depreciar_activo(fecha, request)
                activo.save(request)
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad"})

        elif action == 'depreciaredificio':
            try:
                fecha = convertir_fecha(request.POST['fecha'])
                if Edificio.objects.filter(fechaultimadeprec__gt=fecha, estado=True, depreciable=True).exists():
                    edificio = Edificio.objects.filter(fechaultimadeprec__gt=fecha, estado=True, depreciable=True)[0]
                    return JsonResponse({"result": "bad",
                                         "mensaje": u"Error la fecha de depreciación es incorrecta: La última depreciación tiene fecha %s" % edificio.fechaultimadeprec.strftime(
                                             '%d-%m-%Y')})
                if fecha > datetime.now().date():
                    return JsonResponse({"result": "bad", "mensaje": u"La fecha establecida es mayor a la actual"})
                edificio = Edificio.objects.filter(estado=True, depreciable=True).exclude(vidautil=0)
                edificios = []
                for ac in edificio:
                    act = {'descripcion': unicode(
                        ac.identificacion.descripcion if len(ac.identificacion.descripcion) > 0 else "None"),
                           'id': ac.id}
                    edificios.append(act)
                return JsonResponse({"result": "ok", "cantidad": len(edificios), "edificios": edificios})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al generar lista para depreciar"})

        elif action == 'depreciandoedificio':
            try:
                fecha = convertir_fecha(request.POST['fecha'])
                edificio = Edificio.objects.get(pk=request.POST['id'])
                edificio.depreciar_edificio(fecha, request)
                edificio.save(request)
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad"})

        elif action == 'manten':
            try:
                detalle = int(request.POST['id'])
                mantenimiento = DetalleTrasladoMantenimiento.objects.get(pk=detalle)
                nombre = mantenimiento.observacion
                return JsonResponse({"result": "ok", "nombre": nombre})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'detalle_activo':
            try:
                data['activo'] = activo = ActivoFijo.objects.get(pk=int(request.POST['id']))
                template = get_template("inventario_activofijo/detalle.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'html': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'detalle_activohist':
            try:
                data['activo'] = activo = ActivoFijo.objects.get(id=int(request.POST['id']))
                template = get_template("inventario_activofijo/detalle.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'html': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'detalle_mantenimiento':
            try:
                data['activofijo'] = activofijo = ActivoFijo.objects.get(pk=int(request.POST['id']))
                data['detallemantenimiento'] = HdDetalle_Incidente.objects.filter(incidente__activo=activofijo,
                                                                                  status=True, incidente__status=True,
                                                                                  estado__id=3).order_by(
                    'incidente__fechareporte')
                sgarantia = activofijo.mantenimientosactivospreventivos_set.filter(status=True).order_by()
                data['mantenimientopreventivo'] = sgarantia.order_by('fecha')
                mantenimientogarantia = activofijo.mantenimientosactivosgarantia_set.filter(status=True)
                data['mantenimientogarantia'] = mantenimientogarantia.order_by('fechainicio')
                data['costomantenimientogarantia'] = \
                mantenimientogarantia.filter(status=True).aggregate(cantidad=Sum('valor'))['cantidad']
                template = get_template("inventario_activofijo/detallehelpdesk.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'html': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'detalle_control':
            try:
                tarjeta = TarjetaControl.objects.get(pk=int(request.POST['id']))
                data['mantenimiento'] = tarjeta.detallemantenimiento_set.all().order_by('-fecha_creacion')[0]
                template = get_template("inventario_activofijo/detalle_control.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'html': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'detalle_constatacion':
            try:
                data['constatacion'] = constatacion = ConstatacionFisica.objects.get(pk=int(request.POST['id']))
                data['detallepert'] = detalles = constatacion.detalleconstatacionfisica_set.filter(
                    perteneceusuario=True)
                data['detallenopert'] = detalles = constatacion.detalleconstatacionfisica_set.filter(
                    perteneceusuario=False)
                data['detallenoiden'] = detalles = constatacion.detallenoidentificado_set.all()
                template = get_template("inventario_activofijo/detallecons.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'html': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'detalle_bajas':
            try:
                data['baja'] = baja = BajaActivo.objects.get(pk=int(request.POST['id']))
                data['detalles'] = baja.detallebajaactivo_set.all()
                template = get_template("inventario_activofijo/detallebaja.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'html': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'detalle_movimiento':
            try:
                data['traspaso'] = traspaso = TraspasoActivo.objects.get(pk=int(request.POST['id']))
                data['detalles'] = traspaso.detalletraspasoactivo_set.filter(seleccionado=True)
                template = get_template("inventario_activofijo/detalletras.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'html': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'detalle_actas':
            try:
                data['traspaso'] = traspaso = TraspasoActivo.objects.get(pk=int(request.POST['id']))
                data['detalles'] = traspaso.detalletraspasoactivo_set.filter(codigotraspaso__tipo=1)
                template = get_template("inventario_activofijo/detalleactas.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'html': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'detalle_traslados':
            try:
                data['traslado'] = traslado = TrasladoMantenimiento.objects.get(pk=int(request.POST['id']))
                data['detalles'] = traslado.detalletrasladomantenimiento_set.filter(seleccionado=True)
                template = get_template("inventario_activofijo/detallemantenimiento.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'html': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'usuariosubicacion':
            try:
                ubicacion = Ubicacion.objects.get(pk=int(request.POST['id']))
                lista = []
                for usuario in Persona.objects.filter(responsableactivo__ubicacion=ubicacion,
                                                      responsableactivo__statusactivo=1).distinct():
                    if [usuario.id, usuario.nombre_completo_inverso()] not in lista:
                        lista.append([usuario.id, usuario.nombre_completo_inverso()])
                return JsonResponse({"result": "ok", "lista": lista})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'ubicacionusuario':
            try:
                usuario = Persona.objects.get(pk=int(request.POST['id']))
                lista = []
                for ubicacion in Ubicacion.objects.filter(activofijo__responsable=usuario,
                                                          activofijo__statusactivo=1).distinct():
                    if [ubicacion.id, ubicacion.nombre] not in lista:
                        lista.append([ubicacion.id, ubicacion.nombre])
                return JsonResponse({"result": "ok", "lista": lista})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'usuariocustodio':
            try:
                responsable = Persona.objects.get(pk=int(request.POST['id']))
                ubicacion = Ubicacion.objects.get(pk=int(request.POST['idp']))
                lista = []
                for custodio in Persona.objects.filter(custodioactivo__ubicacion=ubicacion,
                                                       custodioactivo__responsable=responsable,
                                                       custodioactivo__statusactivo=1).distinct():
                    if [custodio.id, custodio.nombre_completo_inverso()] not in lista:
                        lista.append([custodio.id, custodio.nombre_completo_inverso()])
                return JsonResponse({"result": "ok", "lista": lista})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'custodioentrega':
            try:
                custodio = Persona.objects.get(pk=int(request.POST['id']))
                lista = []
                for ubicacion in Ubicacion.objects.filter(activofijo__custodio=custodio,
                                                          activofijo__statusactivo=1).distinct():
                    if [ubicacion.id, ubicacion.nombre] not in lista:
                        lista.append([ubicacion.id, ubicacion.nombre])
                return JsonResponse({"result": "ok", "lista": lista})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'consult_cat':
            try:
                catalogo = CatalogoBien.objects.get(pk=int(request.POST['id']))
                lista = []
                for responsable in Persona.objects.filter(responsableactivo__catalogo=catalogo).distinct():
                    if [responsable.id, responsable.nombre_completo_inverso()] not in lista:
                        lista.append([responsable.id, responsable.nombre_completo_inverso()])
                return JsonResponse({"result": "ok", "lista": lista})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'activosusuariocustodio':
            try:
                custodio = Persona.objects.get(pk=int(request.POST['id']))
                ubicacion = Ubicacion.objects.get(pk=int(request.POST['idp']))
                responsable = Persona.objects.get(pk=int(request.POST['idu']))
                data = {}
                data['activos'] = responsable.mis_activos_ubicacion_custodio(ubicacion, custodio, responsable)
                template = get_template("inventario_activofijo/activospersonaubicacion.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'data': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'activosusuariocustodiotraspaso':
            try:
                custodio = Persona.objects.get(pk=int(request.POST['id']))
                ubicacion = Ubicacion.objects.get(pk=int(request.POST['idp']))
                responsable = Persona.objects.get(pk=int(request.POST['idu']))
                data = {}
                data['activos'] = responsable.mis_activos_ubicacion_custodio(ubicacion, custodio, responsable)
                template = get_template("inventario_activofijo/activosusuariocustodiotraspaso.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'data': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'activos_catalogo':
            try:
                responsable = Persona.objects.get(pk=int(request.POST['id']))
                fechai = convertir_fecha(request.POST['fi'])
                fechaf = convertir_fecha(request.POST['ff'])
                catalogo = CatalogoBien.objects.get(pk=int(request.POST['idc']))
                data = {}
                data['activos'] = ActivoFijo.objects.filter(responsable=responsable, catalogo=catalogo,
                                                            fechaingreso__range=(fechai, fechaf), statusactivo=1)
                data['reporte_0'] = obtener_reporte('consulta_activos_catalogo')
                template = get_template("inventario_activofijo/activoscatalogo.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'data': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'activos_ubicacion':
            try:
                ubicacion = Ubicacion.objects.get(pk=int(request.POST['id']))
                responsable = Persona.objects.get(pk=int(request.POST['idr']))
                data = {}
                data['activos'] = ActivoFijo.objects.filter(responsable=responsable, ubicacion=ubicacion,
                                                            statusactivo=1)
                template = get_template("inventario_activofijo/activosubicacion.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'data': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'activosbaja':
            try:
                responsable = Persona.objects.get(pk=int(request.POST['idp']))
                ubicacion = Ubicacion.objects.get(pk=int(request.POST['id']))
                data = {}
                data['activos'] = responsable.mis_activos_ubicacion(ubicacion)
                template = get_template("inventario_activofijo/activosbaja.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'data': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'activocaracteristica':
            try:
                catalogo = CatalogoBien.objects.get(pk=int(request.POST['id']))
                return JsonResponse({"result": "ok", "tipo": catalogo.tipobien.id})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'misactivos':
            try:
                responsable = Persona.objects.get(pk=int(request.POST['idp']))
                ubicacion = Ubicacion.objects.get(pk=int(request.POST['id']))
                data = {}
                data['activos'] = responsable.mis_activos_ubicacion(ubicacion)
                template = get_template("inventario_activofijo/activosconstatacion.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'data': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'activosdeotros':
            try:
                data = {}
                activo = None
                json_content = ""
                if ActivoFijo.objects.filter(
                        Q(codigogobierno=request.POST['id']) | Q(codigointerno=request.POST['id'])).exists():
                    data['activo'] = activo = ActivoFijo.objects.filter(
                        Q(codigogobierno=request.POST['id']) | Q(codigointerno=request.POST['id'])).distinct()[0]
                    template = get_template("inventario_activofijo/activosotros.html")
                    json_content = template.render(data)
                return JsonResponse({"result": "ok", "cantidad": 1 if activo else 0, 'data': json_content,
                                     "id": activo.id if activo else 0})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'activoubicacion':
            try:
                responsable = Persona.objects.get(pk=int(request.POST['id']))
                lista = []
                for ubicacion in Ubicacion.objects.filter(activofijo__responsable=responsable,
                                                          activofijo__statusactivo=1).distinct():
                    if [ubicacion.id, ubicacion.nombre] not in lista:
                        lista.append([ubicacion.id, ubicacion.nombre])
                return JsonResponse({"result": "ok", "lista": lista})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'addestadobien':
            try:
                form = EstadoBienForm(request.POST)
                if form.is_valid():
                    estado = EstadoBien(nombre=form.cleaned_data['nombre'],
                                        descripcion=form.cleaned_data['descripcion'])
                    estado.save(request)
                    log(u'Adicionó estado de Bienes Inmuebles de la Intitución: %s' % estado, request, "add")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.rollback()
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'editestadobien':
            try:
                form = EstadoBienForm(request.POST)
                if form.is_valid():
                    if not EstadoBien.objects.get(pk=int(request.POST['id'])).esta_activo():
                        estado = EstadoBien.objects.get(pk=int(request.POST['id']))
                        estado.nombre = form.cleaned_data['nombre']
                        estado.descripcion = form.cleaned_data['descripcion']
                        estado.save(request)
                        log(u'Editó estado de Bienes Inmuebles de la Intitución: %s' % estado, request, "edit")
                        return JsonResponse({"result": "ok"})
                    else:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"El Estado de Inmuebles y Bienes esta activo"})
            except Exception as ex:
                transaction.rollback()
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'delestadobien':
            try:
                if not EstadoBien.objects.get(pk=int(request.POST['id'])).esta_activo():
                    estado = EstadoBien.objects.get(pk=int(request.POST['id']))
                    log(u'Eliminó estado de Bienes Inmuebles de la Intitución: %s' % estado, request, "del")
                    estado.delete()
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.rollback()
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'addcondicionbien':
            try:
                form = CondicionBienForm(request.POST)
                if form.is_valid():
                    condicion = CondicionBien(nombre=form.cleaned_data['nombre'],
                                              descripcion=form.cleaned_data['descripcion'])
                    condicion.save(request)
                    log(u'Adicionó Condición de Bienes Inmuebles de la Intitución: %s' % condicion, request, "add")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.rollback()
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'editcondicionbien':
            try:
                form = CondicionBienForm(request.POST)
                if form.is_valid():
                    if not CondicionBien.objects.get(pk=int(request.POST['id'])).esta_activo():
                        condicion = EstadoBien.objects.get(pk=int(request.POST['id']))
                        condicion.nombre = form.cleaned_data['nombre']
                        condicion.descripcion = form.cleaned_data['descripcion']
                        condicion.save(request)
                        log(u'Editó Condición de Bienes Inmuebles de la Intitución: %s' % condicion, request, "edit")
                        return JsonResponse({"result": "ok"})
                    else:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"El Estado de Inmuebles y Bienes esta activo"})
            except Exception as ex:
                transaction.rollback()
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'delcondicionbien':
            try:
                if not CondicionBien.objects.get(pk=int(request.POST['id'])).esta_activo():
                    condicion = CondicionBien.objects.get(pk=int(request.POST['id']))
                    log(u'Eliminó Condición de Bienes Inmuebles de la Intitución: %s' % condicion, request, "del")
                    condicion.delete()
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.rollback()
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'addedificio':
            try:
                form = EdificioForm(request.POST)
                if form.is_valid():
                    edificio = Edificio(codigobien=form.cleaned_data['codigobien'],
                                        codigoanterior=form.cleaned_data['codigoanterior'],
                                        identificador=form.cleaned_data['identificador'],
                                        fechaingreso=form.cleaned_data['fechaingreso'],
                                        catalogo_id=request.POST['catalogo'],
                                        identificacion=form.cleaned_data['identificacion'],
                                        caracteristica=form.cleaned_data['caracteristica'],
                                        propietario=form.cleaned_data['propietario'],
                                        critico=form.cleaned_data['critico'],
                                        clavecatastral=form.cleaned_data['clavecatastral'],
                                        numeropredio=form.cleaned_data['numeropredio'],
                                        numeropiso=form.cleaned_data['numeropiso'],
                                        areaconstruccion=form.cleaned_data['areaconstruccion'],
                                        numeroescritura=form.cleaned_data['numeroescritura'],
                                        fechaescritura=form.cleaned_data['fechaescritura'],
                                        notaria=form.cleaned_data['notaria'],
                                        condicionbien_id=request.POST['condicionbien'],
                                        estadobien_id=request.POST['estadobien'],
                                        cuentacontable_id=request.POST['cuentacontable'],
                                        responsable_id=request.POST['responsable'],
                                        custodio_id=request.POST['custodio'],
                                        vidautil=form.cleaned_data['vidautil'],
                                        depreciable=form.cleaned_data['depreciable'],
                                        valorcontable=form.cleaned_data['valorcontable'],
                                        observacion=form.cleaned_data['observacion'])
                    edificio.save(request)
                    log(u'Adicionó Bienes Inmuebles de la Intitución: %s' % edificio, request, "add")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.rollback()
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'addmantenimientogarantia':
            try:
                form = MantenimientosActivosGarantiaForm(request.POST, request.FILES)
                newfile = None
                if 'archivo' in request.FILES:
                    newfile = request.FILES['archivo']
                    if newfile:
                        newfilesd = newfile._name
                        ext = newfilesd[newfilesd.rfind("."):]
                        if not ext.lower() == '.pdf':
                            return JsonResponse({"result": "bad", "mensaje": u"Error, solo archivos .pdf."})
                        if newfile.size > 2194304:
                            return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 2 Mb."})
                        if newfile:
                            newfile._name = generar_nombre("archivo_malla_", newfile._name)
                if 'arcusen' in request.FILES:
                    newfile = request.FILES['arcusen']
                    if newfile:
                        newfilesd = newfile._name
                        ext = newfilesd[newfilesd.rfind("."):]
                        if not ext.lower() == '.pdf':
                            return JsonResponse({"result": "bad", "mensaje": u"Error, solo archivos .pdf."})
                        if newfile.size > 2194304:
                            return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 2 Mb."})
                        if newfile:
                            newfile._name = generar_nombre("archivo_malla_", newfile._name)
                if form.is_valid():
                    limp = request.POST['id_limp']
                    danio = request.POST['id_dani']
                    id_cpu = int(request.POST['cpu'])
                    # listadoactivos = json.loads(request.POST['lista_items4'])
                    if form.cleaned_data['proveedor'] > 0:
                        proveedor_id = form.cleaned_data['proveedor']
                    else:
                        proveedor_id = None
                    mantenimientogarantia = MantenimientosActivosGarantia(proveedor_id=proveedor_id,
                                                                          valor=form.cleaned_data['valor'],
                                                                          numreporte=form.cleaned_data['numreporte'],
                                                                          fechainicio=form.cleaned_data['fechainicio'],
                                                                          #
                                                                          tipoactivo=form.cleaned_data['tipoactivo'],
                                                                          estusu=form.cleaned_data['estusu'],
                                                                          horamax=form.cleaned_data['horamax'],
                                                                          minutomax=form.cleaned_data['minutomax'],
                                                                          estfrec=form.cleaned_data['estfrec'],
                                                                          estfent=form.cleaned_data['estfent'],
                                                                          observacion=form.cleaned_data['observacion'],
                                                                          activofijo_id=id_cpu,
                                                                          estnuevo=True
                                                                          )
                    mantenimientogarantia.save(request)
                    elementos = limp.split(',')
                    for elemento in elementos:
                        detalle = TareasActivosPreventivosGarantiaLimp(mantenimiento=mantenimientogarantia,
                                                                       grupos_id=elemento)
                        detalle.save()
                    elementosd = danio.split(',')
                    for elemento in elementosd:
                        detalle = TareasActivosPreventivosGarantiaErr(mantenimiento=mantenimientogarantia,
                                                                      grupos_id=elemento)
                        detalle.save()
                    if 'archivo' in request.FILES:
                        newfile = request.FILES['archivo']
                        if newfile:
                            newfile._name = generar_nombre("archivo_garantia", newfile._name)
                            mantenimientogarantia.archivo = newfile
                            mantenimientogarantia.save(request)
                    if 'arcusen' in request.FILES:
                        newfile = request.FILES['arcusen']
                        if newfile:
                            newfile._name = generar_nombre("archivo_entusuario", newfile._name)
                            mantenimientogarantia.arcusen = newfile
                            mantenimientogarantia.save(request)
                    # for codactivo in listadoactivos:
                    #     detallegarantia = DetalleMantenimientosActivosGarantia(mantenimientoactivosgarantia=mantenimientogarantia,
                    #                                                            activofijo_id=codactivo)
                    #     detallegarantia.save(request)
                    log(u'Adicionó mantenimiento de garantia: %s' % mantenimientogarantia, request, "add")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.rollback()
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'editedificio':
            try:
                form = EdificioForm(request.POST)
                if form.is_valid():
                    edificio = Edificio.objects.get(pk=int(request.POST['id']))
                    edificio.codigobien = request.POST['codigobien']
                    edificio.codigoanterior = request.POST['codigoanterior']
                    edificio.identificador = request.POST['identificador']
                    edificio.fechaingreso = form.cleaned_data['fechaingreso']
                    edificio.catalogo_id = request.POST['catalogo']
                    edificio.identificacion_id = request.POST['identificacion']
                    edificio.caracteristica = form.cleaned_data['caracteristica']
                    edificio.propietario = form.cleaned_data['propietario']
                    edificio.critico = form.cleaned_data['critico']
                    edificio.clavecatastral = request.POST['clavecatastral']
                    edificio.numeropredio = str(request.POST['numeropredio'])
                    edificio.numeropiso = request.POST['numeropiso']
                    edificio.areaconstruccion = request.POST['areaconstruccion']
                    edificio.numeroescritura = str(request.POST['numeroescritura'])
                    edificio.fechaescritura = form.cleaned_data['fechaescritura']
                    edificio.notaria = form.cleaned_data['notaria']
                    edificio.condicionbien_id = request.POST['condicionbien']
                    edificio.estadobien_id = request.POST['estadobien']
                    edificio.cuentacontable_id = request.POST['cuentacontable']
                    edificio.responsable_id = request.POST['responsable']
                    edificio.custodio_id = request.POST['custodio']
                    edificio.vidautil = int(request.POST['vidautil'])
                    edificio.depreciable = form.cleaned_data['depreciable']
                    edificio.valorcontable = form.cleaned_data['valorcontable']
                    edificio.observacion = form.cleaned_data['observacion']
                    edificio.save(request)
                    log(u'Editó Bienes Inmuebles de la Intitución: %s' % edificio, request, "edit")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.rollback()
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'deledificio':
            try:
                edificio = Edificio.objects.get(pk=int(request.POST['id']))
                log(u'Eliminó Bienes Inmuebles de la Intitución: %s  el usuario %s' % (edificio, usuario), request,
                    "del")
                edificio.delete()
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.rollback()
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'deltaremantenimiento':
            try:
                edificio = MantenimientoGruCategoria.objects.get(pk=int(request.POST['id']))
                log(u'Eliminó tarea de mantenimiento: %s  el usuario %s' % (edificio, usuario), request, "del")
                edificio.delete()
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.rollback()
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'addtareamantenimiento':
            try:
                form = TareaMantenimientoFrom(request.POST)
                if form.is_valid():
                    edificio = MantenimientoGruCategoria(grupocategoria=form.cleaned_data['categoria'],
                                                         descripcion=form.cleaned_data['descripcion'])
                    edificio.save(request)
                    log(u'Adicionó tarea de mantenimiento: %s' % edificio, request, "add")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.rollback()
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'addtareamantenimientosgdanios':
            try:
                form = TareaMantenimientoDaniosForm(request.POST)
                if form.is_valid():
                    edificio = MantenimientoGruDanios(grupocategoria=form.cleaned_data['categoria'],
                                                      descripcion=(form.cleaned_data['descripcion']).upper())
                    edificio.save(request)
                    log(u'Adicionó tarea de mantenimiento sin garantia en danios: %s' % edificio, request, "add")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.rollback()
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'edittareamantenimientosgdanio':
            try:
                form = TareaMantenimientoFrom(request.POST)
                if form.is_valid():
                    tarea = MantenimientoGruDanios.objects.get(pk=int(request.POST['id']))
                    tarea.grupocategoria = form.cleaned_data['categoria']
                    tarea.descripcion = form.cleaned_data['descripcion'].upper()
                    tarea.save(request)
                    log(u'Aditó tarea de mantenimiento sin garantia en danio: %s' % tarea, request, "edit")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.rollback()
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'deltaremantenimientosgdanio':
            try:
                edificio = MantenimientoGruDanios.objects.get(pk=int(request.POST['id']))
                log(u'Eliminó tarea de mantenimiento sin garantia en danio: %s  el usuario %s' % (edificio, usuario),
                    request, "del")
                edificio.delete()
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.rollback()
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'act_des_tareasgdanio':
            try:
                if 'id' in request.POST:
                    tarea = MantenimientoGruDanios.objects.get(status=True, pk=int(request.POST['id']))
                    if tarea.activo:
                        tarea.activo = False
                    else:
                        tarea.activo = True
                    tarea.save(request)
                    return JsonResponse({"result": "ok", "valor": 1 if tarea.activo else 0})
                else:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al activar la tarea."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        # Mantenimiento de Limpieza
        elif action == 'addtareamantenimientolimpieza':
            try:
                form = TareaMantenimientoFrom(request.POST)
                if form.is_valid():
                    limpieza = MantenimientoGruCategoriaGarantiaLimp(grupocategoria=form.cleaned_data['categoria'],
                                                                     descripcion=form.cleaned_data[
                                                                         'descripcion'].upper())
                    limpieza.save(request)
                    log(u'Adicionó tarea de mantenimiento de limpieza: %s' % limpieza, request, "add")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.rollback()
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'deltaremantenimientolimpieza':
            try:
                limpieza = MantenimientoGruCategoriaGarantiaLimp.objects.get(pk=int(request.POST['id']))
                log(u'Eliminó tarea de mantenimiento en limpieza: %s  el usuario %s' % (limpieza, usuario), request,
                    "del")
                limpieza.delete()
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.rollback()
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'act_des_tarea_limp':
            try:
                if 'id' in request.POST:
                    tarea = MantenimientoGruCategoriaGarantiaLimp.objects.get(status=True, pk=int(request.POST['id']))
                    if tarea.activo:
                        tarea.activo = False
                    else:
                        tarea.activo = True
                    tarea.save(request)
                    return JsonResponse({"result": "ok", "valor": 1 if tarea.activo else 0})
                else:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al activar la tarea."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'edittareamantenimientolimpieza':
            try:
                form = TareaMantenimientoFrom(request.POST)
                if form.is_valid():
                    tarea = MantenimientoGruCategoriaGarantiaLimp.objects.get(pk=int(request.POST['id']))
                    tarea.grupocategoria = form.cleaned_data['categoria']
                    tarea.descripcion = form.cleaned_data['descripcion'].upper()
                    tarea.save(request)
                    log(u'Editó tarea de mantenimiento en limpieza: %s' % tarea, request, "edit")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.rollback()
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'listatipomantenimientolimpieza':
            try:
                listatareas = MantenimientoGruCategoriaGarantiaLimp.objects.filter(
                    grupocategoria_id=request.POST['id_tipo'], status=True, activo=True)
                lista = []
                for lis in listatareas:
                    lista.append([lis.id, lis.descripcion])
                return JsonResponse({'result': 'ok', 'lista': lista})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        # Mantenimientos en daños
        elif action == 'addtareamantenimientodanio':
            try:
                form = TareaMantenimientoFrom(request.POST)
                if form.is_valid():
                    limpieza = MantenimientoGruCategoriaGarantiaErr(grupocategoria=form.cleaned_data['categoria'],
                                                                    descripcion=form.cleaned_data[
                                                                        'descripcion'].upper())
                    limpieza.save(request)
                    log(u'Adicionó tarea de mantenimiento de danio: %s' % limpieza, request, "add")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.rollback()
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'act_des_tarea_danio':
            try:
                if 'id' in request.POST:
                    tarea = MantenimientoGruCategoriaGarantiaErr.objects.get(status=True, pk=int(request.POST['id']))
                    if tarea.activo:
                        tarea.activo = False
                    else:
                        tarea.activo = True
                    tarea.save(request)
                    return JsonResponse({"result": "ok", "valor": 1 if tarea.activo else 0})
                else:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al activar la tarea."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'tareasmantenimientodanio':
            try:
                danios = MantenimientoGruCategoriaGarantiaErr.objects.get(pk=int(request.POST['id']))
                log(u'Eliminó tarea de mantenimiento en danios: %s  el usuario %s' % (danios, usuario), request, "del")
                danios.delete()
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.rollback()
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'edittaremantenimientodanio':
            try:
                form = TareaMantenimientoFrom(request.POST)
                if form.is_valid():
                    tarea = MantenimientoGruCategoriaGarantiaErr.objects.get(pk=int(request.POST['id']))
                    tarea.grupocategoria = form.cleaned_data['categoria']
                    tarea.descripcion = form.cleaned_data['descripcion'].upper()
                    tarea.save(request)
                    log(u'Editó tarea de mantenimiento en danios: %s' % tarea, request, "edit")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.rollback()
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'listatipomantenimientodanio':
            try:
                listatareas = MantenimientoGruCategoriaGarantiaErr.objects.filter(
                    grupocategoria_id=request.POST['id_tipo'], status=True, activo=True)
                lista = []
                for lis in listatareas:
                    lista.append([lis.id, lis.descripcion])
                return JsonResponse({'result': 'ok', 'lista': lista})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'delmantenimiento':
            try:
                mantenimiento = MantenimientosActivosPreventivos.objects.get(pk=int(request.POST['id']))
                log(u'Eliminó mantenimiento: %s  el usuario %s' % (mantenimiento, usuario), request, "del")
                mantenimiento.delete()
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.rollback()
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'delmantenimientogarantia':
            try:
                mantenimientogarantia = MantenimientosActivosGarantia.objects.get(pk=int(request.POST['id']))
                log(u'Eliminó mantenimiento garantía: %s  el usuario %s' % (mantenimientogarantia, usuario), request,
                    "del")
                mantenimientogarantia.delete()
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.rollback()
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'validar_conf':
            try:
                mens = ''
                if not CondicionBien.objects.filter(status=True).exists():
                    mens = u"Condición del Bien, "
                if not EstadoBien.objects.filter(status=True).exists():
                    mens += u"Estado del Bien, "
                if CondicionBien.objects.filter(status=True).exists() and EstadoBien.objects.filter(
                        status=True).exists():
                    return JsonResponse({"result": "ok"})
                else:
                    return JsonResponse(
                        {"result": "bad", "mensaje": "No Tiene Registro Ingresado en " + mens[:mens.__len__() - 2]})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'detalle_edificio':
            try:
                data['edificio'] = Edificio.objects.get(pk=int(request.POST['id']))
                template = get_template("inventario_activofijo/detalleedificio.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'html': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'detalle_mantenimientopreven':
            try:
                data['mantenimiento'] = mantenimiento = MantenimientosActivosPreventivos.objects.get(
                    pk=request.POST['id'])
                # data['tareas']  = TareasActivosPreventivos.objects.filter(mantenimiento=mantenimiento)
                data['tareasmantenimiento'] = MantenimientoGruCategoria.objects.filter(
                    grupocategoria=mantenimiento.tipoactivo, status=True)
                data['tareasactivo'] = TareasActivosPreventivos.objects.values_list('grupos_id', flat=True).filter(
                    mantenimiento=mantenimiento, status=True)
                data['ppmantenimiento'] = HdPiezaPartes.objects.filter(grupocategoria=mantenimiento.tipoactivo,
                                                                       status=True)
                data['piezaparteactivo'] = PiezaParteActivosPreventivos.objects.filter(mantenimiento=mantenimiento,
                                                                                       status=True)
                data['daniomantenimiento'] = MantenimientoGruDanios.objects.filter(
                    grupocategoria=mantenimiento.tipoactivo, status=True)
                data['danioactivo'] = TareasActivosPreventivosDanios.objects.filter(mantenimiento=mantenimiento,
                                                                                    status=True)
                data['estdan'] = ESTADO_DANIO
                if mantenimiento.nuevo:
                    template = get_template("inventario_activofijo/detallemantenimientopreventivov2.html")
                else:
                    template = get_template("inventario_activofijo/detallemantenimientopreventivo.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'html': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'detalle_incidentetec':
            try:
                data['mantenimiento'] = mantenimiento = HdDetalle_Incidente.objects.get(id=request.POST['id'])
                template = get_template("inventario_activofijo/detalleincidentetecnologico.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'html': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'detalle_mantgarantia':
            try:
                data['mantenimiento'] = mantenimiento = MantenimientosActivosGarantia.objects.get(pk=request.POST['id'])
                data['tareaslimpieza'] = MantenimientoGruCategoriaGarantiaLimp.objects.filter(
                    grupocategoria=mantenimiento.tipoactivo, status=True)
                data['tareasdanio'] = MantenimientoGruCategoriaGarantiaErr.objects.filter(
                    grupocategoria=mantenimiento.tipoactivo, status=True)
                data['tareaslimpiezat'] = TareasActivosPreventivosGarantiaLimp.objects.values_list('grupos_id',
                                                                                                   flat=True).filter(
                    mantenimiento=mantenimiento, status=True)
                data['tareasdaniot'] = TareasActivosPreventivosGarantiaErr.objects.values_list('grupos_id',
                                                                                               flat=True).filter(
                    mantenimiento=mantenimiento, status=True)
                template = get_template("inventario_activofijo/detallemantenimientogarantia.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'html': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'addarchivobaja':
            try:
                form = ArchivoActivoBajaForm(request.POST, request.FILES)
                if form.is_valid():
                    activo = ActivoFijo.objects.get(pk=request.POST['id'])
                    newfile = request.FILES['archivobaja']
                    newfile._name = generar_nombre("evidencia_", newfile._name)
                    if activo.archivobaja:
                        log(u'Edito archivo de activo dado de baja en Activo Fijo: activo = %s - [%s], archivo antiguo = %s, archivo nuevo = %s' % (
                        activo, activo.id, activo.archivobaja, newfile), request, "edit")
                    else:
                        log(u'Adiciono archivo de activo dado de baja en Activo Fijo: activo = %s - [%s], archivo = %s' % (
                        activo, activo.id, newfile), request, "add")
                    activo.archivobaja = newfile
                    activo.save(request)
                    return JsonResponse({"result": False}, safe=False)
                else:
                    raise NameError('Debe subir archivo pdf con los MG especificado')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'act_des_tarea':
            try:
                if 'id' in request.POST:
                    tarea = MantenimientoGruCategoria.objects.get(status=True, pk=int(request.POST['id']))
                    if tarea.activo:
                        tarea.activo = False
                    else:
                        tarea.activo = True
                    tarea.save(request)
                    return JsonResponse({"result": "ok", "valor": 1 if tarea.activo else 0})
                else:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al activar la tarea."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'edittareamantenimiento':
            try:
                form = TareaMantenimientoFrom(request.POST)
                if form.is_valid():
                    tarea = MantenimientoGruCategoria.objects.get(pk=int(request.POST['id']))
                    tarea.grupocategoria = form.cleaned_data['categoria']
                    tarea.descripcion = form.cleaned_data['descripcion']
                    tarea.save(request)
                    log(u'Aditó tarea de mantenimiento: %s' % tarea, request, "edit")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.rollback()
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'addmanactivofijo':
            try:
                codigoactivofijo = request.POST['codigoactivofijo']
                mangarantiaid = request.POST['mangarantiaid']
                if DetalleMantenimientosActivosGarantia.objects.filter(mantenimientoactivosgarantia_id=mangarantiaid,
                                                                       activofijo_id=codigoactivofijo).exists():
                    return JsonResponse({"result": "bad", "mensaje": u"Activo ya se encuentra registrado."})
                detalle = DetalleMantenimientosActivosGarantia(mantenimientoactivosgarantia_id=mangarantiaid,
                                                               activofijo_id=codigoactivofijo)
                detalle.save(request)
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'delmanactivofijo':
            try:
                detalle = DetalleMantenimientosActivosGarantia.objects.get(pk=int(request.POST['hidencodigodetalle']))
                detalle.delete()
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        if action == 'addActivoSinCod':
            try:
                f = ActivoFijoForm(request.POST)
                if f.is_valid():
                    if f.cleaned_data['responsable'] or f.cleaned_data['ubicacion']:
                        if not f.cleaned_data['responsable'] or not f.cleaned_data['ubicacion']:
                            return JsonResponse(
                                {"result": "bad", "mensaje": u"Debe seleccionar Responsable y Ubicación."})
                    activo = ActivoFijo(
                        nombre=f.cleaned_data['nombre'],
                        descripcion=f.cleaned_data['descripcion'],
                        serie=f.cleaned_data['serie'],
                        modelo=f.cleaned_data['modelo'],
                        marca=f.cleaned_data['marca'],
                        vidautil=f.cleaned_data['vidautil'],
                        estado=f.cleaned_data['estado'])
                    activo.save(request)
                    if f.cleaned_data['responsable']:
                        activo.ubicacion = f.cleaned_data["ubicacion"]
                        activo.responsable_id = f.cleaned_data["responsable"]

                    log(u'Adicionó Activo: %s' % activo, request, "add")
                    cuenta = CuentaContable.objects.get(pk=int(activo.cuentacontable_id))
                    cuenta.activosfijos = True
                    cuenta.save(request)
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})


        elif action == 'addgrupo':
            try:
                catalogo = CatalogoBien.objects.get(pk=int(request.POST['id']))
                f = GrupoBienForm(request.POST)
                if f.is_valid():
                    catalogo.grupo = f.cleaned_data['grupo']
                    catalogo.save(request)
                    log(u'Adicionó un grupo a catálogo: %s' % catalogo, request, "add")
                    return JsonResponse({"result": False}, safe=False)
                else:
                    raise NameError('Error al enviar los datos')

            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})


        elif action == 'addresponsable':
            try:

                form = ResponsableActivoForm(request.POST)
                if form.is_valid():
                    responsable = form.cleaned_data['persona']
                    activo = int(request.POST['activo'])
                    verificarsolicitud = SolicitudActivos.objects.filter(status=True, activo_id=activo,
                                                                         responsableasignacion_id=responsable.id,
                                                                         estado=1)
                    verificarresponsable = ActivoTecnologico.objects.filter(status=True, responsable_id=responsable.id,
                                                                     pk=activo)
                    if not verificarsolicitud:
                        if not verificarresponsable:
                            solicitud = SolicitudActivos(activo_id=activo, responsableasignacion_id=responsable.id,
                                                         solicitante=personaactivo,
                                                         fechasolicitud=datetime.now().date())
                            solicitud.save(request)
                            log(u'Envio de solicitud traspaso: %s' % solicitud, request, "add")
                            return JsonResponse({"result": False}, safe=False)
                        return JsonResponse({"result": True, "mensaje": "El activo ya cuenta con el mismo responsable"},
                                            safe=False)
                    return JsonResponse({"result": True, "mensaje": "El activo mantiene una solicitud pendiente"},
                                        safe=False)
                else:
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                         "mensaje": "Error en el formulario"})
            except Exception as ex:
                pass

        elif action == 'addinventario':
            try:
                form = InventarioActivoForm(request.POST)
                if form.is_valid():
                    observacion = form.cleaned_data['observacion']
                    activo = int(request.POST['activo'])
                    fecha = datetime.now().date()
                    observacion = form.cleaned_data['observacion']
                    estadoubicacion = form.cleaned_data['estadoubicacion']
                    estadofuncionamiento = form.cleaned_data['estadofuncionamiento']
                    estadouso = form.cleaned_data['estadouso']
                    estadogeneral = request.POST.getlist('estadogeneral')

                    periodoexistente = PeriodoInventarioAT.objects.filter(status=True, estado=1)
                    if not periodoexistente:
                        return JsonResponse({"result": True, "mensaje": f"No existe un periodo vigente para el inventario actual"})
                    existencia = ActivoFijoInventarioTecnologico.objects.filter((Q(status=True) & Q(activo_id=activo)) & (Q(periodo__estado=1) | Q(periodo_id__isnull=True)))
                    if existencia.exists():
                        return JsonResponse({"result": True, "mensaje": f"El activo ya se encuentra en inventario"})

                    if fecha > periodoexistente[0].fechafin:
                        return JsonResponse({"result": True, "mensaje": f"El periodo del inventario ha finalizado"})

                    inventario = ActivoFijoInventarioTecnologico(activo_id=activo, observacion=observacion,
                                                                 fechainventario=datetime.now(),
                                                                 periodo_id=periodoexistente[0].id,
                                                                 estadoubicacion=estadoubicacion,
                                                                 estadofuncionamiento=estadofuncionamiento)
                    inventario.save(request)
                    log(u'Adicionó activo en inventario: %s' % inventario, request, "add")
                    for estadoat in estadogeneral:
                        estadogeneralat = InventarioATEstadosGenerales(inventarioat_id=inventario.id,estadogeneral_id=estadoat)
                        estadogeneralat.save(request)
                        log(u'Adiciono estado general en registro inventario: %s' % estadogeneralat, request, "add")
                    return JsonResponse({"result": False, "mensaje": f"Activo registrado correctamente"}, safe=False)
                else:
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                         "mensaje": "Error en el formulario"})
            except Exception as ex:
                pass


        elif action == 'solicitudeliminar':
            try:
                id = int(request.POST['id'])
                solicitud = SolicitudActivos.objects.get(pk=id, status=True)
                solicitud.estado = 6
                solicitud.save(request)
                log(u'Solicitante elimina solicitud: %s' % solicitud, request, "del")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                pass

        elif action == 'eliminaperiodo':
            try:
                id = int(request.POST['id'])
                # periodoenuso = ActivoFijoInventarioTecnologico.objects.filter(status=True, periodo_id=id)
                # if not periodoenuso:
                periodoeliminar = PeriodoInventarioAT.objects.get(status=True, pk=id)
                if periodoeliminar.total_cronograma() != 0:
                    periodoeliminar.status = False
                    periodoeliminar.save(request)
                    log(u'Eliminación de periodo: %s' % periodoeliminar, request, "del")
                else:
                    raise NameError('Periodo en uso no pude ser eliinado.')
                res_json = {"error": False}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)


        elif action == 'generarcodigotics':
            try:
                mayor = 0
                codigoinicial = 1
                codigosiguiente = 0
                contadorasignado = 0
                activosmigrados = ActivoTecnologico.objects.filter(status=True)
                if not activosmigrados:
                    return JsonResponse({"result": "ok", "contador": contadorasignado, "migrados": "No"})
                verificar = ActivoTecnologico.objects.filter(status=True).exclude(codigotic=None)
                if verificar:
                    mayor = ActivoTecnologico.objects.filter(status=True).aggregate(
                        codigotic=Max(Cast(('codigotic'), IntegerField())))['codigotic']
                    codigosiguiente = int(mayor) + 1
                    activos = ActivoTecnologico.objects.filter(status=True, codigotic=None).order_by('id')
                    if activos:
                        for activo in activos:
                            activo.codigotic = codigosiguiente
                            activo.save(request)
                            codigosiguiente = codigosiguiente + 1
                            contadorasignado = contadorasignado + 1
                            log(u'Creación de códigotic: %s' % activo, request, "add")
                else:
                    activos = ActivoTecnologico.objects.filter(status=True, codigotic=None).order_by('id')
                    if activos:
                        for activo in activos:
                            activo.codigotic = codigoinicial
                            activo.save(request)
                            codigoinicial = codigoinicial + 1
                            contadorasignado = contadorasignado + 1

                return JsonResponse({"result": "ok", "contador": contadorasignado})
            except Exception as ex:
                pass


        elif action == 'migraractivos':
            try:
                codigoinicial = 1
                contadorasignado = 0
                codigosiguiente = 1
                mayor = 0
                activosfijos = ActivoFijo.objects.filter(status=True)
                for activo in activosfijos:
                    activoexistente = ActivoTecnologico.objects.filter(status=True, activotecnologico_id=activo.id)
                    if not activoexistente:
                        tecnologico = ActivoTecnologico(activotecnologico_id=activo.id)
                        tecnologico.save(request)
                        contadorasignado = contadorasignado + 1
                        log(u'Adicionó activo en Activo Tecnológico: %s' % tecnologico, request, "add")
                return JsonResponse({"result": "ok", "contador": contadorasignado})
            except Exception as ex:
                pass


        elif action == 'generarcodigoqrtic':
            try:
                id = int(request.POST['id'])
                activo = ActivoTecnologico.objects.get(activotecnologico_id=id, status=True)
                ruta = 'https://sga.unemi.edu.ec/inventario_activofijo?cod=' + str(activo.codigotic)
                codigoactivo = str(activo.codigotic)
                rutaqr = SITE_ROOT + '/media/qrcode/activos/Activo' + str(activo.codigotic) + '.png'
                url = pyqrcode.create(ruta)
                # url.svg(rutaqr, scale=4)
                url.png('media/qrcode/activos/Activo' + str(activo.codigotic) + '.png', 10, '#000000')
                return JsonResponse({"result": "ok", "ruta": rutaqr, "activo": codigoactivo})
            except Exception as ex:
                pass

        elif action == 'adicionarinventario':
            try:
                form = InventarioActivoForm(request.POST)
                if form.is_valid():
                    fecha = datetime.now().date()
                    activo = int(request.POST['activotecnologico'])
                    # movimiento = form.cleaned_data['movimiento']
                    observacion = form.cleaned_data['observacion']
                    estadoubicacion = form.cleaned_data['estadoubicacion']
                    estadofuncionamiento = form.cleaned_data['estadofuncionamiento']
                    estadogeneral = request.POST.getlist('estadogeneral')
                    periodoexistente = PeriodoInventarioAT.objects.filter(status=True, estado=1)
                    if not periodoexistente:
                        return JsonResponse({"result": True, "mensaje": "No existe un periodo vigente para el inventario actual"})
                    existencia = ActivoFijoInventarioTecnologico.objects.filter((Q(status=True) & Q(activo_id=activo)) & (Q(periodo__estado=1) | Q(periodo_id__isnull=True)))
                    if existencia.exists():
                        return JsonResponse({"result":True, "mensaje":"El activo ya se encuentra en inventario"})
                    if fecha <= periodoexistente[0].fechafin:
                        inventario = ActivoFijoInventarioTecnologico(activotecnologico_id=activo, observacion=observacion,
                                                                     fechainventario=datetime.now(), periodo_id=periodoexistente[0].id,
                                                                     estadoubicacion=estadoubicacion,
                                                                     estadofuncionamiento=estadofuncionamiento)
                        inventario.save(request)
                        for estadoat in estadogeneral:
                            estadogeneralat = InventarioATEstadosGenerales(inventarioat_id=inventario.id, estadogeneral_id=estadoat)
                            estadogeneralat.save(request)
                            log(u'Adiciono estado general en registro inventario: %s' % estadogeneralat, request, "add")
                        log(u'Adiciono activo en inventario: %s' % inventario, request, "add")
                        return JsonResponse({"result": False}, safe=False)
                    return JsonResponse({"result": True, "mensaje": "El periodo del inventario ha finalizado"})
                else:
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                         "mensaje": "Error en el formulario"})
            except Exception as ex:
                return JsonResponse({"result": True, "mensaje": "Por favor, llene el formulario correctamente"})

        elif action == 'adicionarestadogeneral':
            try:
                form = EstadosGeneralesInventarioATForm(request.POST)
                if form.is_valid():
                    descripcion = form.cleaned_data['descripcion']
                    estadogeneral = EstadosGeneralesInventarioAT(descripcion=descripcion)
                    estadogeneral.save(request)
                    log(u'Adiciono estado general: %s' % estadogeneral, request, "add")
                    return JsonResponse({"result": False}, safe=False)
                else:
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                         "mensaje": "Error en el formulario"})
            except Exception as ex:
                return JsonResponse({"result": True, "mensaje": "Por favor, llene el formulario correctamente"})


        elif action == 'editarregistroinventario':
            try:
                form = InventarioActivoForm(request.POST)
                if form.is_valid():
                    activo = int(request.POST['activo'])
                    observacion = form.cleaned_data['observacion']
                    estadoubicacion = form.cleaned_data['estadoubicacion']
                    estadogeneral = request.POST.getlist('estadogeneral')
                    estadofuncionamiento = form.cleaned_data['estadofuncionamiento']
                    estadouso = form.cleaned_data['estadouso']
                    # movimiento = form.cleaned_data['movimiento']
                    existencia = ActivoFijoInventarioTecnologico.objects.get(status=True, pk=activo)
                    existencia.observacion = observacion
                    existencia.estadoubicacion = estadoubicacion
                    existencia.estadofuncionamiento = estadofuncionamiento
                    existencia.estadouso = estadouso
                    # existencia.movimiento = movimiento
                    existencia.save(request)
                    registroestados = []
                    registroestadogeneral = InventarioATEstadosGenerales.objects.filter(status=True, estadogeneral_id__in=estadogeneral,
                                                                                        inventarioat_id=existencia.id)
                    compararestados = InventarioATEstadosGenerales.objects.filter(inventarioat_id=existencia.id,
                                                                                    status=True).exclude(estadogeneral_id__in=estadogeneral)
                    if len(estadogeneral) == 0:
                        registroestge = InventarioATEstadosGenerales.objects.filter(status=True,inventarioat_id=existencia.id)
                        for registro in registroestge:
                            registro.status = False
                            registro.save(request)
                    else:
                        for estadogen in estadogeneral:
                            registroestge = InventarioATEstadosGenerales.objects.filter(status=True,
                                                                                        estadogeneral_id=estadogen,
                                                                                        inventarioat_id=existencia.id)
                            if not registroestge:
                                estadogeneralat = InventarioATEstadosGenerales(inventarioat_id=existencia.id,
                                                                               estadogeneral_id=estadogen)
                                estadogeneralat.save(request)
                                log(u'Adiciono estado general en registro inventario: %s' % estadogeneralat, request,
                                    "add")
                    for estadogen in compararestados:
                        estadogen.status = False
                        estadogen.save(request)
                    log(u'Actualizó observación en inventario: %s' % existencia, request, "act")
                    return JsonResponse({"result": False, "mensaje":u'Registro actualizado correctamente!'}, safe=False)
                else:
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                         "mensaje": "Error en el formulario"})
            except Exception as ex:
                return JsonResponse({'result': True, "mensaje":u'Por favor, llene el formulario correctamente.'})


        elif action == 'cargarcombo_activo':
            try:
                activos = ActivoFijo.objects.filter((Q(archivobaja__isnull=True) |
                                                     Q(archivobaja='')) &
                                                    Q(catalogo__equipoelectronico=True) &
                                                    Q(catalogo__status=True) &
                                                    Q(statusactivo=1) &
                                                    Q(status=True))
                lista = []
                if activos:
                    for activo in activos:
                        lista.append([activo.id, activo.descripcion.__str__(), activo.codigointerno.__str__(),
                                      activo.codigogobierno.__str__()])
                return JsonResponse({'result': 'ok', 'lista': lista})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})


        elif action == 'cargarcombo_responsable':
            try:
                responsables = Persona.objects.filter(Q(status=True) & (
                                                 Q(perfilusuario__administrativo__isnull=False) |
                                                 Q(perfilusuario__profesor__isnull=False))).distinct().order_by('apellido1',
                                                                                                                'apellido2',
                                                                                                                'nombres')
                lista = []
                if responsables:
                    for responsable in responsables:
                        lista.append([responsable.id, responsable.nombres.__str__(), responsable.apellido1.__str__(),
                                      responsable.apellido2.__str__(), responsable.identificacion()])
                return JsonResponse({'result': 'ok', 'lista': lista})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})


        elif action == 'reporteactivoinventario':
            try:
                if 'activotecnologico' in request.POST:
                    activo = int(request.POST['activotecnologico'])
                    data['listadoactivos'] = inventario = ActivoFijoInventarioTecnologico.objects.filter(status=True,
                                                                                            activo_id=activo)
                    if not inventario:
                        data['activofijo'] = ActivoFijo.objects.filter(status=True, pk=activo)
                    data['fecha'] = datetime.now().date()
                    data['usuariogenera'] = usuario.username
                    data['estado'] = True
                    return conviert_html_to_pdf('inventario_activofijo/reporteinventarioactivo.html',
                                                {'pagesize': 'A4', 'data': data})
            except Exception as ex:
                pass

        elif action == 'reporteresponsableactivoinventario':
            try:
                if 'activotecnologico' in request.POST and 'responsable' in request.POST:
                    activo = int(request.POST['activotecnologico'])
                    responsable = int(request.POST['responsable'])
                    data['responsable'] = Persona.objects.filter(status=True, pk=responsable)
                    data['listadoactivos'] = inventario = ActivoFijoInventarioTecnologico.objects.filter(status=True,
                                                                                            activo_id=activo, activo__responsable_id=responsable)

                    if not inventario:
                        data['activofijo'] = ActivoFijo.objects.filter(status=True,pk=activo)
                    data['fecha'] = datetime.now().date()
                    data['estado'] = True
                else:
                    if 'responsable' in request.POST:
                        responsable = int(request.POST['responsable'])
                        data['fecha'] = datetime.now().date()
                        data['estado'] = False
                        data['responsable'] = Persona.objects.filter(status=True, pk=responsable)
                        data['listadoactivos'] = inventario = ActivoFijoInventarioTecnologico.objects.filter(status=True,
                                                                                                      activo__responsable_id=responsable)
                data['usuariogenera'] = usuario.username
                return conviert_html_to_pdf('inventario_activofijo/reporteinventarioactivo.html',
                                            {'pagesize': 'A4', 'data': data})
            except Exception as ex:
                pass

        elif action == 'addperiodoinventario':
            try:
                form = PeriodoInventarioATForm(request.POST)
                if form.is_valid():
                    nombre = request.POST['nombre']
                    fechainicio = request.POST['fechainicio']
                    fechafin = request.POST['fechafin']
                    detalle = request.POST['detalle']
                    estado = form.cleaned_data['estado']
                    periodovigente = PeriodoInventarioAT.objects.filter(status=True, estado=1)
                    if not periodovigente:
                        if fechafin >= fechainicio:
                            periodoinventario = PeriodoInventarioAT(nombre=nombre, fechainicio=fechainicio,
                                                                    estado=estado,
                                                                    fechafin=fechafin, detalle=detalle)
                            periodoinventario.save(request)
                            log(u'Adicionó un nuevo periodo: %s' % periodoinventario, request, "add")
                            return JsonResponse({"result": False}, safe=False)
                        return JsonResponse({"result": True, "mensaje": "La fecha de finalización del periodo tiene que ser mayor o igual a la fecha de inicio"})
                    return JsonResponse({"result": True, "mensaje": "Ya existe un periodo vigente"})
                else:
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                         "mensaje": "Error en el formulario"})
            except Exception as ex:
                pass


        elif action == 'editperiodoinventario':
            try:
                form = PeriodoInventarioATForm(request.POST)
                if form.is_valid():
                    periodo = None
                    nombre = request.POST['nombre']
                    fechainicio = request.POST['fechainicio']
                    fechafin = request.POST['fechafin']
                    detalle = request.POST['detalle']
                    if 'idperiodo' in request.POST:
                        periodo = int(request.POST['idperiodo'])
                        if fechafin >= fechainicio:
                            periodovigente = PeriodoInventarioAT.objects.get(status=True, pk=periodo)
                            periodovigente.nombre = nombre
                            periodovigente.fechainicio = fechainicio
                            periodovigente.fechafin = fechafin
                            periodovigente.estado = form.cleaned_data['estado']
                            periodovigente.detalle = detalle
                            periodovigente.save(request)
                            log(u'Editó el periodo: %s' % periodovigente, request, "act")
                            return JsonResponse({"result": False}, safe=False)
                        return JsonResponse({"result": True,"mensaje": "La fecha de finalización del periodo tiene que ser mayor o igual a la fecha de inicio"})
                    return JsonResponse({"result": True, "mensaje": "No se ha podido realizar la actualización"})
                else:
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                         "mensaje": "Error en el formulario"})
            except Exception as ex:
                pass


        elif action == 'asignarperiodoeninventario':
            try:
                periodoexistente = PeriodoInventarioAT.objects.filter(status=True, estado=1)
                if not periodoexistente:
                    return JsonResponse({"result":"bad", "mensaje":"No existe un periodo vigente"})

                registroinventario = ActivoFijoInventarioTecnologico.objects.filter(status=True, periodo_id__isnull=True)
                if registroinventario:
                    if periodoexistente:
                        for inventario in registroinventario:
                            inventario.periodo_id = periodoexistente[0].id
                            inventario.save(request)
                            log(u'Adicinó periodo en inventario: %s' % inventario, request, "act")
                        # asignacion = AsignacionCierreInventarioAT(persona=personaactivo,
                        #                                           periodo_id=periodoexistente[0].id,
                        #                                           fecha=datetime.now().date(), estado=1)
                        # asignacion.save(request)
                        # log(u'Asignó periodo en inventario: %s' % asignacion, request, "act")
                        return JsonResponse({"result": "ok"})
                else:
                    return JsonResponse({"result": "bad", "mensaje": "No existen registros en inventario"})
            except Exception as ex:
                pass


        elif action == 'cierreinventario':
            try:
                registroinventario = ActivoFijoInventarioTecnologico.objects.filter(status=True, periodo_id__isnull=True)
                if not registroinventario:
                    fecha = datetime.now().date()
                    periodoinventario = PeriodoInventarioAT.objects.filter(status=True, estado=1)
                    if periodoinventario:
                        if fecha > periodoinventario[0].fechafin:
                            listaregistrosinventario = []
                            correousuario = []
                            informeactivosinventario = ''
                            responsableseninventario = Persona.objects.filter(Q(status=True) & (
                                                     Q(perfilusuario__administrativo__isnull=False) |
                                                     Q(perfilusuario__profesor__isnull=False))).distinct()
                            for responsableinv in responsableseninventario:
                                correousuario.clear()
                                correousuario.append(responsableinv.emailinst)
                                listaregistrosinventario.clear()
                                inventics = ActivoFijoInventarioTecnologico.objects.filter(status=True, activotecnologico__responsable=responsableinv, periodo_id=int(request.POST['idperiodoinv']))
                                if inventics:
                                    for registroinv in inventics:
                                        listaregistrosinventario.append(registroinv)
                                    nombreinformeinv = 'activostecnologicosinv' + responsableinv.identificacion()
                                    valida = conviert_html_to_pdfsaveinformeinventarioactivostecnologicos(
                                        'inventario_activofijo/activostecnoeninventarioporresponsable.html',
                                        {'pagesize': 'A4', 'data': data, 'registroseninventario': listaregistrosinventario,
                                         'fecha': datetime.now().date()},
                                        nombreinformeinv + '.pdf')

                                    if valida:
                                        informeactivosinventario = os.path.join(SITE_STORAGE, 'media',
                                                                      'inventarioactivostecnologicos/' + nombreinformeinv + '.pdf')
                                    asunto = u"Activos Tecnológicos en Inventario"
                                    send_html_mail(asunto, "emails/activostecnologicoseninventariotics.html",
                                                   {'sistema': 'SGA UNEMI', 'responsable': responsableinv},
                                                   correousuario, [], [informeactivosinventario],
                                                   cuenta=CUENTAS_CORREOS[1][1])
                            periodoinventario[0].fechacierre = datetime.now().date()
                            periodoinventario[0].estado = 2
                            asignacion = AsignacionCierreInventarioAT(persona=personaactivo, periodo_id=periodoinventario[0].id,
                                                                      fecha=datetime.now().date(), estado=1)
                            asignacion.save(request)
                            periodoinventario[0].save(request)
                            log(u'Realizó cierre inventario: %s' % periodoinventario[0], request, "act")
                            return JsonResponse({"result":"ok"})
                        return JsonResponse({"result": "bad", "mensaje": "El periodo del inventario aún no finaliza"})
                    return JsonResponse({"result": "bad", "mensaje": "No existe periodo vigente"})
                else:
                    return JsonResponse({"result": "bad", "mensaje": "Registros en inventario no cuentan con periodo vigente"})
            except Exception as ex:
                pass



        elif action == 'reporte_certificadoprevio':
            try:
                data['activo'] = ActivoFijo.objects.filter(status=True, pk=int(request.POST['id']))
                qrname = 'qr_certificado_'
                return conviert_html_to_pdfsave(
                    'inventario_activofijo/notificacionentregaactivo.html',
                    {'pagesize': 'A4', 'data': data}, qrname + '.pdf'
                )
            except Exception as ex:
                pass

        elif action == 'subirevidenciaperiodo':
            try:
                f = EvidenciaPeriodoInventarioTecnologicoForm(request.POST, request.FILES)
                if f.is_valid():
                    if 'evidencia' in request.FILES:
                        arch = request.FILES['evidencia']
                        extension = arch._name.split('.')
                        tam = len(extension)
                        exte = extension[tam - 1]
                        if arch.size > 4194304:
                            return JsonResponse(
                                {"result": True, "mensaje": u"Error, el tamaño del archivo es mayor a 4 Mb."})
                        if not exte.lower() == 'pdf':
                            return JsonResponse({"result": True, "mensaje": u"Solo se permiten archivos .pdf"})
                        newfile = request.FILES['evidencia']
                        newfile._name = generar_nombre("evidenciaperiodoIT", newfile._name)
                        # periodoinvat = PeriodoInventarioAT.objects.get(status=True, estado=1)
                        evidencia = EvidenciaPeriodoInventarioTecnologico(nombre=f.cleaned_data['nombre'],
                                                                          descripcion=f.cleaned_data['descripcion'],
                                                                          periodo_id=int(request.POST['idperiodo']),
                                                                          fecha=datetime.now().date(),
                                                                          archivo=newfile)
                        evidencia.save(request)
                        log(u'Añade evidencia del periodo: %s' % evidencia, request, "add")
                        return JsonResponse({'result': False}, safe=False)
                else:
                    for k, v in f.errors.items():
                        raise NameError(v[0])
            except Exception as ex:
                pass

        elif action == 'editarevidenciaperiodo':
            try:
                f = EvidenciaPeriodoInventarioTecnologicoForm(request.POST, request.FILES)
                archivoevidencia = request.FILES.getlist('evidencia')
                newfile = None
                evidencia = EvidenciaPeriodoInventarioTecnologico.objects.get(status=True, id=int(request.POST['id']))
                if archivoevidencia:
                    if f.is_valid():
                        if 'evidencia' in request.FILES:
                            arch = request.FILES['evidencia']
                            extension = arch._name.split('.')
                            tam = len(extension)
                            exte = extension[tam - 1]
                            if arch.size > 4194304:
                                return JsonResponse(
                                    {"result": "bad", "mensaje": u"Error, el tamaño del archivo es mayor a 4 Mb."})
                            if not exte.lower() == 'pdf':
                                return JsonResponse({"result": "bad", "mensaje": u"Solo se permiten archivos .pdf"})
                            newfile = request.FILES['evidencia']
                            newfile._name = generar_nombre("evidenciaperiodoIT", newfile._name)
                    else:
                        for k, v in f.errors.items():
                            raise NameError(v[0])
                if not newfile:
                    newfile = evidencia.archivo
                evidencia.nombre = request.POST['nombre']
                evidencia.descripcion = request.POST['descripcion']
                evidencia.archivo = newfile
                evidencia.save(request)
                log(u'Actualiza evidencia del periodo: %s' % evidencia, request, "act")
                return JsonResponse({'result': False, 'mensaje': u'Actualización exitosa'})
            except Exception as ex:
                pass

        elif action == 'eliminarevidenciaperiodoinv':
           try:
               # periodoinv = PeriodoInventarioAT.objects.get(status=True, id=int(request.POST['id']))
               # idevidenciaperiodo = periodoinv.evidencia_id
               # periodoinv.evidencia = None
               evidenciaperiodo = EvidenciaPeriodoInventarioTecnologico.objects.get(status=True, id=int(request.POST['id']))
               evidenciaperiodo.status = False
               # periodoinv.save(request)
               evidenciaperiodo.save(request)
               log(u'Elimina evidencia de periodo: %s' % evidenciaperiodo, request, "del")
               return JsonResponse({"result": "ok", "mensaje": u'Evidencia eliminada correctamente!'})
           except Exception as ex:
               return JsonResponse({"result": False, "mensaje":u'Ha ocurrido un error'})

        elif action == 'eliminarestadogeneral':
           try:
               id = int(request.POST['id'])
               estadogeneral = EstadosGeneralesInventarioAT.objects.get(id=id, status=True)
               estadogeneral.status = False
               estadogeneral.save(request)
               log(u'Eliminó estado general: %s' % estadogeneral, request, "del")
               return JsonResponse({"result": "ok"})
           except Exception as ex:
               pass

        elif action == 'editarestadogeneral':
            try:
                form = EstadosGeneralesInventarioATForm(request.POST)
                if form.is_valid():
                    idestadogeneral = int(request.POST['id'])
                    descripcion = form.cleaned_data['descripcion']
                    estadogeneralinventarioat = EstadosGeneralesInventarioAT.objects.get(status=True, id=idestadogeneral)
                    estadogeneralinventarioat.descripcion = descripcion
                    estadogeneralinventarioat.save(request)
                    log(u'Actualizó estado general: %s' % estadogeneralinventarioat, request, "act")
                    return JsonResponse({"result": False}, safe=False)
                else:
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                         "mensaje": "Error en el formulario"})
            except Exception as ex:
                pass

        #Cronograma
        elif action == 'addevento':
            try:
                idp=int(encrypt(request.POST['idp']))
                form = CronogramaPersonaInventarioForm(request.POST)
                if form.is_valid() and form.validador(0,idp):
                    cronograma=CronogramaPersonaConstatacionAT.objects.filter(status=True, periodo_id=idp)
                    numero = len(cronograma)+1
                    instancia = CronogramaPersonaConstatacionAT(periodo_id=idp,
                                                                numero=numero,
                                                                descripcion=form.cleaned_data['descripcion'],
                                                                persona=form.cleaned_data['persona'],
                                                                fecha=form.cleaned_data['fecha'],
                                                                hora=form.cleaned_data['hora'])
                    instancia.save(request)

                    ids_activos_t=instancia.activos_tecnologicos().values('id')
                    for id_activo_t in ids_activos_t:
                        detalle=DetalleConstatacionFisicaActivoTecnologico(cronograma_id=instancia.id, activo_id=id_activo_t['id'])
                        detalle.save(request)
                else:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                         "mensaje": "Error en el formulario"})
                log(u'Adiciono Evento: %s' % instancia, request, "add")
                return JsonResponse({"result": False}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, "mensaje": 'Error: {}'.format(ex)}, safe=False)

        elif action == 'editevento':
            try:
                id = int(encrypt(request.POST['id']))
                instancia=CronogramaPersonaConstatacionAT.objects.get(id=id)
                if instancia.estado >= 3 :
                    raise NameError('Cronograma no puede ser modificado por que el registro ya fue finalizado o cerrado.')
                form = CronogramaPersonaInventarioForm(request.POST)
                if form.is_valid() and form.validador(id,instancia.periodo.id):
                    instancia.descripcion=form.cleaned_data['descripcion']
                    instancia.persona=form.cleaned_data['persona']
                    instancia.fecha=form.cleaned_data['fecha']
                    instancia.hora=form.cleaned_data['hora']
                    instancia.save(request)
                else:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                         "mensaje": "Error en el formulario"})
                log(u'Edito Evento: %s' % instancia, request, "edit")
                return JsonResponse({"result": False}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, "mensaje": 'Error: {}'.format(ex)}, safe=False)

        elif action == 'editfecha':
            try:
                id = int(request.POST['id'])
                fecha = datetime.strptime(request.POST['val'], '%d/%m/%Y')
                instancia=CronogramaPersonaConstatacionAT.objects.get(id=id)
                if instancia.fechainicio:
                    raise NameError('Constatación ya fue iniciada por lo que no se pude cambiar de fecha de este evento')
                instancia.fecha=fecha.date()
                instancia.save(request)
                log(u'Edito Evento: %s' % instancia, request, "edit")
                return JsonResponse({"result": True, "mensaje": 'Guardado con exito'}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': False, "mensaje": 'Error: {}'.format(ex)}, safe=False)

        elif action == 'delevento':
            try:
                id = int(request.POST['id'])
                instancia=CronogramaPersonaConstatacionAT.objects.get(id=id)
                if not instancia.puede_eliminar():
                    raise NameError("Evento que intenta eliminar ya tiene constataciones realizadas.")
                instancia.status=False
                instancia.save(request)
                log(u'Eliminó cronograma: %s' % instancia, request, "del")
                res_json = {"error": False}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        elif action == 'inicioconstatacion':
            try:
                id = int(request.POST['id'])
                instancia = CronogramaPersonaConstatacionAT.objects.get(id=id)
                instancia.fechainicio = hoy
                instancia.responsable_id = personaactivo.id
                instancia.estado = 2
                instancia.save(request)
                # ids_activos_t = instancia.activos_tecnologicos().values('id')
                # for id_activo_t in ids_activos_t:
                #     detalle = DetalleConstatacionFisicaActivoTecnologico(cronograma_id=instancia.id,
                #                                                          activo_id=id_activo_t['id'])
                #     detalle.save(request)
                log(u'Edito fecha inicio constatación cronograma: %s' % instancia, request, "edit")
                res_json = {"result": True}
            except Exception as ex:
                res_json = {'result': False, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        #Constataciones
        elif action == 'addconstatacionat':
            try:
                idcronograma=int(encrypt(request.POST['idcronograma']))
                idactivo=int(encrypt(request.POST['idactivo']))
                if DetalleConstatacionFisicaActivoTecnologico.objects.filter(cronograma_id=idcronograma, status=True, activo_id=idactivo):
                    raise NameError('Constatación de activo tecnologico ya existe.')
                aplica = True if 'aplica' in request.POST else False
                activo_t=ActivoTecnologico.objects.get(id=idactivo)
                form = ConstatacionFisicaATForm(request.POST)
                if not 'encontrado' in request.POST:
                    form.fields['estadoactual'].required=False
                if form.is_valid():
                    instancia = DetalleConstatacionFisicaActivoTecnologico(cronograma_id=idcronograma,
                                                                            activo_id=idactivo,
                                                                            encontrado=form.cleaned_data['encontrado'],
                                                                            enuso = form.cleaned_data['enuso'],
                                                                            perteneceusuario = form.cleaned_data['perteneceusuario'],
                                                                            # ubicacionbienes = form.cleaned_data['ubicacionbienes'],
                                                                            bloque = form.cleaned_data['bloque'],
                                                                            ubicacion = form.cleaned_data['ubicacion'],
                                                                            estadoactual = form.cleaned_data['estadoactual'],
                                                                            estadooriginal = activo_t.activotecnologico.estado,
                                                                            usuariobienes = form.cleaned_data['usuariobienes'],
                                                                            requieretraspaso = form.cleaned_data['requieretraspaso'],
                                                                            requieredarbaja = form.cleaned_data['requieredarbaja'],
                                                                            aplica = aplica,
                                                                            observacion = form.cleaned_data['observacion'])
                    instancia.save(request)
                    if aplica:
                        activo = ActivoTecnologico.objects.get(id=idactivo)
                        componentes=ComponenteCatalogoActivo.objects.filter(status=True,catalogo_id=activo.activotecnologico.catalogo.id)
                        for componente in componentes:
                            estado = int(request.POST[f'est_{componente.nombre_input()}'])
                            observacion = request.POST[f'obs_{componente.nombre_input()}']
                            encontrado = True if f'encontrado_{componente.nombre_input()}' in request.POST else False
                            catalogo=DetalleCatalogoComponenteConstatacionAT(constatacion=instancia,
                                                                            catalogo=componente,
                                                                             estado_id=estado,
                                                                             encontrado=encontrado,
                                                                             observacion=observacion)
                            catalogo.save(request)
                else:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                         "mensaje": "Error en el formulario"})
                log(u'Adiciono Constatación Fisica AT: %s' % instancia, request, "add")
                return JsonResponse({"result": False}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, "mensaje": 'Error: {}'.format(ex)}, safe=False)

        elif action == 'editconstatacionat':
            try:
                id = int(encrypt(request.POST['id']))
                aplica = True if 'aplica' in request.POST else False
                instancia = DetalleConstatacionFisicaActivoTecnologico.objects.get(id=id)
                form = ConstatacionFisicaATForm(request.POST)
                if not 'encontrado' in request.POST:
                    form.fields['estadoactual'].required=False
                if form.is_valid():
                    instancia.encontrado = form.cleaned_data['encontrado']
                    instancia.enuso = form.cleaned_data['enuso']
                    instancia.perteneceusuario = form.cleaned_data['perteneceusuario']
                    # instancia.ubicacionbienes = form.cleaned_data['ubicacionbienes']
                    instancia.bloque = form.cleaned_data['bloque']
                    instancia.ubicacion = form.cleaned_data['ubicacion']
                    instancia.estadoactual = form.cleaned_data['estadoactual']
                    instancia.estadooriginal = instancia.activo.activotecnologico.estado
                    instancia.usuariobienes = form.cleaned_data['usuariobienes']
                    instancia.requieretraspaso = form.cleaned_data['requieretraspaso']
                    instancia.requieredarbaja = form.cleaned_data['requieredarbaja']
                    instancia.aplica = aplica
                    instancia.observacion = form.cleaned_data['observacion']
                    instancia.constatado = True
                    instancia.save(request)
                    if aplica:
                        componentes = ComponenteCatalogoActivo.objects.filter(status=True, catalogo_id=instancia.activo.activotecnologico.catalogo.id)
                        for componente in componentes:
                            detalle_c=instancia.detalle_componentes().filter(catalogo_id=componente.id).first()
                            if detalle_c:
                                detalle_c.estado_id=int(request.POST[f'est_{detalle_c.nombre_input()}'])
                                detalle_c.observacion=request.POST[f'obs_{detalle_c.nombre_input()}']
                                detalle_c.encontrado= True if f'encontrado_{detalle_c.nombre_input()}' in request.POST else False
                                detalle_c.save(request)
                            else:
                                estado = int(request.POST[f'est_{componente.nombre_input()}'])
                                observacion = request.POST[f'obs_{componente.nombre_input()}']
                                encontrado= True if f'encontrado_{componente.nombre_input()}' in request.POST else False
                                catalogo = DetalleCatalogoComponenteConstatacionAT(constatacion=instancia,
                                                                                   catalogo=componente,
                                                                                   estado_id=estado,
                                                                                   encontrado=encontrado,
                                                                                   observacion=observacion)
                                catalogo.save(request)
                    else:
                        componentes = ComponenteCatalogoActivo.objects.filter(status=True,catalogo_id=instancia.activo.activotecnologico.catalogo.id)
                        for componente in componentes:
                            detalle_c=instancia.detalle_componentes().filter(catalogo_id=componente.id).update(status=False)
                            # if detalle_c:
                            #     detalle_c.status=False
                            #     detalle_c.save(request)

                else:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                         "mensaje": "Error en el formulario"})
                log(u'Edito Constatación Fisica AT: %s' % instancia, request, "edit")
                return JsonResponse({"result": False}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, "mensaje": 'Error: {}'.format(ex)}, safe=False)

        elif action == 'delconstatacionat':
            try:
                id = int(encrypt(request.POST['id']))
                instancia=DetalleConstatacionFisicaActivoTecnologico.objects.get(id=id)
                instancia.status=False
                instancia.save(request)
                log(u'Eliminó constatación fisica AT: %s' % instancia, request, "del")
                res_json = {"error": False}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        elif action == 'prefinalizacion':
            try:
                id = int(encrypt(request.POST['id']))
                instancia = CronogramaPersonaConstatacionAT.objects.get(id=id)
                instancia.responsable_id = personaactivo.id
                instancia.fechafin = hoy
                instancia.estado = 3
                instancia.save(request)

                titulo = u"Finalización de constatación física de activos"
                mensaje= u"Por favor revisar la constatación realizada y confirme que todo esta en orden"
                notificacion(titulo, mensaje, instancia.persona, None,
                             f'/mis_activos?action=constataciones&id={encrypt(instancia.id)}',
                             instancia.pk, 1, 'sga', CronogramaPersonaConstatacionAT, request)
                log(u'Finalizo la constatación de activos tecnologicos: %s' % instancia, request, "edit")
                res_json = {"result": True}
            except Exception as ex:
                res_json = {'result': False, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        elif action == 'notificar':
            try:
                id = int(encrypt(request.POST['id']))
                instancia = PeriodoInventarioAT.objects.get(id=id)
                estados=request.POST.getlist('estado')
                mensaje=request.POST['mensaje']
                finicio=request.POST['fechainicio']
                ffin=request.POST['fechafin']
                filtro=Q(estado__in=estados)
                if finicio:
                    filtro = filtro & (Q(fecha__gte=finicio))
                if ffin:
                    filtro = filtro & (Q(fecha__lte=ffin))
                eventos=instancia.eventos().filter(filtro)
                if not estados:
                    raise NameError('Seleccione como minimo un estado')
                if not eventos:
                    raise NameError('No existen eventos con el estado marcado')
                
                titulo = u"Constatación física de activos tecnológicos"
                for evento in eventos:
                    lista_email = evento.persona.lista_emails()
                    # lista_email = ['walarconr@unemi.edu.ec', ]
                    datos_email = {'sistema': request.session['nombresistema'],
                                   'fecha': datetime.now().date(),
                                   'hora': datetime.now().time(),
                                   'evento': evento,
                                   'persona': evento.persona,
                                   'mensaje': mensaje}
                    template = "emails/notificarconstatacion.html"
                    send_html_mail(titulo, template, datos_email, lista_email, [], [],cuenta=CUENTAS_CORREOS[0][1])
                log(u'Notifico mensaje a personas consideradas en cronograma: %s' % instancia, request, "edit")
                return JsonResponse({"result": False}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, "mensaje": 'Error: {}'.format(ex)}, safe=False)

        elif action == 'notificarpersonalizada':
            try:
                id = int(encrypt(request.POST['id']))
                instancia = PeriodoInventarioAT.objects.get(id=id)
                estados=request.POST.getlist('estado')
                titulo=request.POST['titulo']
                cabecera=request.POST['cabecera']
                mensaje=request.POST['mensaje_at']
                finicio=request.POST['fechainicio']
                ffin=request.POST['fechafin']
                filtro=Q(estado__in=estados)
                if finicio:
                    filtro = filtro & (Q(fecha__gte=finicio))
                if ffin:
                    filtro = filtro & (Q(fecha__lte=ffin))
                eventos=instancia.eventos().filter(filtro)
                if not estados:
                    raise NameError('Seleccione como minimo un estado')
                if not eventos:
                    raise NameError('No existen eventos con el estado marcado')

                noti_fechas = 'notifechas' in request.POST
                for evento in eventos:
                    lista_email = evento.persona.lista_emails()
                    # lista_email = ['walarconr@unemi.edu.ec', ]
                    datos_email = {'sistema': request.session['nombresistema'],
                                   'fecha': datetime.now().date(),
                                   'hora': datetime.now().time(),
                                   'evento': evento,
                                   'titulo': titulo,
                                   'cabecera': cabecera,
                                   'persona': evento.persona,
                                   'noti_fechas': noti_fechas,
                                   'mensaje': mensaje}
                    template = "emails/notificacionpersonalizada_at.html"
                    send_html_mail(titulo, template, datos_email, lista_email, [], [],cuenta=CUENTAS_CORREOS[0][1])
                log(u'Notifico mensaje a personas consideradas en cronograma: %s' % instancia, request, "edit")
                return JsonResponse({"result": False}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, "mensaje": 'Error: {}'.format(ex)}, safe=False)

        elif action == 'migraractivosconstatacion':
            try:
                id = int(encrypt(request.POST['id']))
                instancia = CronogramaPersonaConstatacionAT.objects.get(id=id)
                constataciones=instancia.constataciones().values_list('activo_id')
                ids_activos_t = instancia.activos_tecnologicos().exclude(id__in=constataciones).values('id')
                for id_activo_t in ids_activos_t:
                    detalle = DetalleConstatacionFisicaActivoTecnologico(cronograma_id=instancia.id,
                                                                         activo_id=id_activo_t['id'])
                    detalle.save(request)
                log(u'Migro activos a constataciónes de activos tecnologicos: %s' % instancia, request, "add")
                res_json = {"result": True}
            except Exception as ex:
                res_json = {'result': False, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        elif action == 'firmaracta':
            try:
                # Parametros
                txtFirmas = json.loads(request.POST['txtFirmas'])
                if not txtFirmas:
                    raise NameError("Debe seleccionar ubicación de la firma")
                x = txtFirmas[-1]
                id_cronograma = int(encrypt(request.POST['id_objeto']))
                responsables = request.POST.getlist('responsables[]')
                firma = request.FILES["firma"]
                passfirma = request.POST['palabraclave']
                url_archivo = (SITE_STORAGE + request.POST["url_archivo"]).replace('\\', '/')
                _name = generar_nombre(f'actafirmada_{request.user.username}_{id_cronograma}_', 'firmada')
                folder = os.path.join(SITE_STORAGE, 'media', 'activostecnologicos/acta_constatacion/', '')

                # Firmar y guardar archivo en folder definido.
                firma = firmararchivogenerado(request, passfirma, firma, url_archivo, folder, _name, x["numPage"], x["x"], x["y"], x["width"], x["height"])
                if firma != True:
                    raise NameError(firma)
                folder_save = os.path.join('activostecnologicos/acta_constatacion/', '').replace('\\', '/')
                url_file_generado = f'{folder_save}{_name}.pdf'
                historial = HistorialDocumentosFirmadosConstatacionAT(cronograma_id=id_cronograma, persona=personaactivo,orden=1)
                historial.save(request)
                historial.archivo = url_file_generado
                historial.save(request)
                log(u'Guardo archivo firmado: {}'.format(historial), request, "add")
                return JsonResponse({"result": False, "mensaje": "Guardado con exito", }, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": u"Error al guardar. %s" % ex.__str__()}, safe=False)

        elif action == 'cerrareventos':
            try:
                idperiodo=int(encrypt(request.POST['id']))
                cronogramas=CronogramaPersonaConstatacionAT.objects.filter(status=True, estado=3, periodo_id=idperiodo)
                for cronograma in cronogramas:
                    cronograma.estado=4
                    cronograma.save(request)
                    titulo = u"Cierre de constatación física de activos tecnológicos"
                    mensaje = u"Se cerro la constataciones de activos tecnológicos"
                    notificacion(titulo, mensaje, cronograma.persona, None,
                                 f'/mis_activos?action=constataciones&id={encrypt(cronograma.id)}',
                                  cronograma.pk, 1, 'sga', CronogramaPersonaConstatacionAT, request)
                    log(u'Edito estado a cerrar : %s' % cronograma, request, "edit")
                return JsonResponse({"result": "ok", "mensaje": "Guardado con exito", }, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar. %s" % ex.__str__()}, safe=False)

        return JsonResponse({"result": "bad", "mensaje": u"Solicitud Incorrecta."})
    else:
        if 'action' in request.GET:
            data['action'] = action = request.GET['action']

            if action == 'estadobien':
                try:
                    data['title'] = u'Estado de Bienes Inmuebles de la Institución'
                    data['estados'] = EstadoBien.objects.filter(status=True)
                    return render(request, "inventario_activofijo/view_estadobien.html", data)
                except:
                    pass

            elif action == 'addestadobien':
                try:
                    data['title'] = u'Adicionar Estado de Bienes Inmuebles de la Institución'
                    data['form'] = EstadoBienForm()
                    return render(request, "inventario_activofijo/addestdobien.html", data)
                except:
                    pass

            elif action == 'editestadobien':
                try:
                    data['title'] = u'Editar Estado de Bienes Inmuebles de la Institución'
                    data['estado'] = estado = EstadoBien.objects.get(pk=int(request.GET['id']))
                    data['form'] = EstadoBienForm(initial={'nombre': estado.nombre, 'descripcion': estado.descripcion})
                    return render(request, "inventario_activofijo/editestadobien.html", data)
                except:
                    pass

            elif action == 'delestadobien':
                try:
                    data['title'] = u'Eliminar Estado de Bienes Inmuebles de la Institución'
                    data['estado'] = EstadoBien.objects.get(pk=int(request.GET['id']))
                    return render(request, "inventario_activofijo/delestadobien.html", data)
                except:
                    pass

            elif action == 'condicionbien':
                try:
                    data['title'] = u'Condiciones de Bienes Inmuebles de la Institución'
                    data['condiciones'] = CondicionBien.objects.filter(status=True)
                    return render(request, "inventario_activofijo/view_condicionbien.html", data)
                except:
                    pass

            elif action == 'addcondicionbien':
                try:
                    data['title'] = u'Adicionar Condición de Bienes Inmuebles de la Institución'
                    data['form'] = CondicionBienForm()
                    return render(request, "inventario_activofijo/addcondicionbien.html", data)
                except:
                    pass

            elif action == 'editcondicionbien':
                try:
                    data['title'] = u'Editar Condición de Bienes Inmuebles de la Institución'
                    data['condicion'] = estado = CondicionBien.objects.get(pk=int(request.GET['id']))
                    data['form'] = CondicionBienForm(
                        initial={'nombre': estado.nombre, 'descripcion': estado.descripcion})
                    return render(request, "inventario_activofijo/editcondicionbien.html", data)
                except:
                    pass

            elif action == 'delcondicionbien':
                try:
                    data['title'] = u'Eliminar Condición de Bienes Inmuebles de la Institución'
                    data['condicion'] = CondicionBien.objects.get(pk=int(request.GET['id']))
                    return render(request, "inventario_activofijo/delcondicionbien.html", data)
                except:
                    pass

            elif action == 'edificios':
                try:
                    data['title'] = u'Bienes Inmuebles de la Institución'
                    # data['edificios'] = Edificio.objects.filter(status=True)
                    search = None
                    ids = None
                    if 's' in request.GET:
                        search = request.GET['s'].strip()
                        ss = search.split(' ')
                        if len(ss) == 1:
                            edificios = Edificio.objects.filter(Q(identificacion__descripcion__icontains=search) |
                                                                Q(codigobien=search) |
                                                                Q(codigoanterior=search) |
                                                                Q(catalogo__descripcion__icontains=search) |
                                                                Q(responsable__nombres__icontains=search) |
                                                                Q(responsable__apellido1__icontains=search) |
                                                                Q(responsable__apellido2__icontains=search) |
                                                                Q(cuentacontable__cuenta__icontains=search) |
                                                                Q(cuentacontable__descripcion__icontains=search) |
                                                                Q(identificador=search) |
                                                                Q(condicionbien__nombre__icontains=search) |
                                                                Q(estadobien__nombre__icontains=search)).distinct().order_by(
                                'fechaingreso')
                        else:
                            edificios = Edificio.objects.filter(Q(responsable__apellido1__icontains=ss[0]) | Q(
                                responsable__apellido2__icontains=ss[1]) |
                                                                (Q(identificacion__descripcion__icontains=ss[0]) & Q(
                                                                    identificacion__descripcion__icontains=ss[
                                                                        1]))).distinct().order_by('fechaingreso')
                    elif 'id' in request.GET:
                        ids = int(request.GET['id'])
                        edificios = Edificio.objects.filter(id=ids)
                    else:
                        edificios = Edificio.objects.filter(status=True).order_by('fechaingreso')
                    paging = MiPaginador(edificios, 15)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['edificios'] = page.object_list
                    data['reporte_0'] = obtener_reporte('edificio_cuenta')
                    data['cuentascontables'] = CuentaContable.objects.filter(status=True)
                    data['form001'] = ReporteEdificioFrom()
                    return render(request, 'inventario_activofijo/view_edificio.html', data)
                except Exception as ex:
                    pass

            elif action == 'mantenimientos':
                try:
                    data['title'] = u'Mantenimientos preventivos sin garantía'
                    search = None
                    ids = None
                    data['tipos'] = GruposCategoria.objects.all()
                    if 's' in request.GET:
                        search = request.GET['s'].strip()
                        ss = search.split(' ')
                        if len(ss) == 1:
                            mantenimientosactivos = MantenimientosActivosPreventivos.objects.filter(
                                Q(activofijo__codigointerno=search) | Q(activofijo__codigogobierno=search),
                                status=True).order_by('fecha')
                        else:
                            mantenimientosactivos = MantenimientosActivosPreventivos.objects.filter(
                                status=True).order_by('-fecha')
                    else:
                        # mantenimientosactivos = MantenimientosActivosPreventivos.objects.filter(status=True).order_by('-fecha')
                        mantenimientosactivos = MantenimientosActivosPreventivos.objects.filter(
                            tipoactivo__descripcion='COMPUTADOR DE ESCRITORIO').order_by('-fecha')
                    if 'tipo' in request.GET:
                        tipo = request.GET['tipo']
                        data['idtipo'] = int(tipo)
                        mantenimientosactivos = MantenimientosActivosPreventivos.objects.filter(
                            tipoactivo__id=tipo).order_by('-fecha')
                    if 'id' in request.GET:
                        ids = request.GET['id']
                        mantenimientosactivos = MantenimientosActivosPreventivos.objects.filter(id=ids).order_by(
                            '-fecha')
                    if ('fini' in request.GET) and ('ffin' in request.GET) and ('tipo' in request.GET):
                        fini = request.GET['fini']
                        ffin = request.GET['ffin']
                        tipo = request.GET['tipo']
                        mantenimientosactivos = MantenimientosActivosPreventivos.objects.filter(tipoactivo__id=tipo,
                                                                                                fecha__range=[fini,
                                                                                                              ffin])

                    paging = MiPaginador(mantenimientosactivos, 15)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['totalact'] = mantenimientosactivos.count()
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['mantenimientosactivos'] = page.object_list
                    return render(request, 'inventario_activofijo/view_mantenimientos.html', data)
                except Exception as ex:
                    pass

            elif action == 'tareasmantenimiento':
                try:
                    data['title'] = u'Tareas de mantenimientos'
                    search = None
                    ids = None
                    if 's' in request.GET:
                        search = request.GET['s'].strip()
                        ss = search.split(' ')
                        if len(ss) == 1:
                            tareasmantenimiento = MantenimientoGruCategoria.objects.filter(
                                descripcion__icontains=search, status=True).order_by('grupocategoria_id', 'descripcion')
                        else:
                            tareasmantenimiento = MantenimientoGruCategoria.objects.filter(
                                descripcion__icontains=search, status=True).order_by('grupocategoria_id', 'descripcion')
                    else:
                        tareasmantenimiento = MantenimientoGruCategoria.objects.filter(status=True).order_by(
                            'grupocategoria_id', 'descripcion')
                    paging = MiPaginador(tareasmantenimiento, 15)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['tareasmantenimiento'] = page.object_list
                    return render(request, 'inventario_activofijo/view_tareasmantenimientos.html', data)
                except Exception as ex:
                    pass

            elif action == 'tareasmantenimientosgdanio':
                try:
                    data['title'] = u'Lista de daños encontrados'
                    search = None
                    ids = None
                    if 's' in request.GET:
                        search = request.GET['s'].strip()
                        ss = search.split(' ')
                        if len(ss) == 1:
                            tareasmantenimiento = MantenimientoGruDanios.objects.filter(
                                Q(descripcion__icontains=search) | Q(grupocategoria__descripcion__icontains=search),
                                status=True).order_by('grupocategoria_id', 'descripcion')
                        else:
                            tareasmantenimiento = MantenimientoGruDanios.objects.filter(descripcion__icontains=search,
                                                                                        status=True).order_by(
                                'grupocategoria_id', 'descripcion')
                    else:
                        tareasmantenimiento = MantenimientoGruDanios.objects.filter(status=True).order_by(
                            'grupocategoria_id', 'descripcion')
                    paging = MiPaginador(tareasmantenimiento, 15)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['tareasmantenimiento'] = page.object_list
                    return render(request, 'inventario_activofijo/view_tareasmantenimientossgdanio.html', data)
                except Exception as ex:
                    pass

            elif action == 'addtareamantenimientosgdanios':
                try:
                    # puede_realizar_accion(request, 'sagest.puede_ingresar_compras')
                    data['title'] = u'Adicionar daño encontrado'
                    data['form'] = TareaMantenimientoDaniosForm()
                    plantilla = render(request, 'inventario_activofijo/addtareamantenimientosgdanios.html', data)
                    return plantilla
                except Exception as ex:
                    pass

            elif action == 'edittareamantenimientosgdanio':
                try:
                    data['title'] = u'Editar daño encontrado'
                    data['tarea'] = tarea = MantenimientoGruDanios.objects.get(status=True, pk=int(request.GET['id']))
                    data['form'] = TareaMantenimientoDaniosForm(
                        initial={'categoria': tarea.grupocategoria, 'descripcion': tarea.descripcion})
                    return render(request, "inventario_activofijo/edittareamantenimientosgdanio.html", data)
                except:
                    pass

            elif action == 'deltaremantenimientosgdanio':
                try:
                    data['title'] = u'Eliminar daño encontrado'
                    data['tareamantenimiento'] = MantenimientoGruDanios.objects.get(pk=int(request.GET['id']))
                    return render(request, "inventario_activofijo/deltareamantenimientosgdanio.html", data)
                except:
                    pass

            elif action == 'tareasmantenimientolimpieza':
                try:
                    data['title'] = u'Lista de tareas en limpieza'
                    search = None
                    ids = None
                    if 's' in request.GET:
                        search = request.GET['s'].strip()
                        ss = search.split(' ')
                        if len(ss) == 1:
                            tareaslimpieza = MantenimientoGruCategoriaGarantiaLimp.objects.filter(
                                descripcion__icontains=search, status=True).order_by('grupocategoria_id', 'descripcion')
                        else:
                            tareaslimpieza = MantenimientoGruCategoriaGarantiaLimp.objects.filter(
                                descripcion__icontains=search, status=True).order_by('grupocategoria_id', 'descripcion')
                    else:
                        tareaslimpieza = MantenimientoGruCategoriaGarantiaLimp.objects.filter(status=True).order_by(
                            'grupocategoria_id', 'descripcion')
                    paging = MiPaginador(tareaslimpieza, 15)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['tareasmantenimiento'] = page.object_list
                    return render(request, 'inventario_activofijo/view_tareasmantenimientoslimpieza.html', data)
                except Exception as ex:
                    pass

            elif action == 'addtareamantenimientolimpieza':
                try:
                    # puede_realizar_accion(request, 'sagest.puede_ingresar_compras')
                    data['title'] = u'Adicionar tarea de mantenimiento en limpieza'
                    data['form'] = TareaMantenimientoFrom()
                    plantilla = render(request, 'inventario_activofijo/addtareamantenimientolimpieza.html', data)
                    return plantilla
                except Exception as ex:
                    pass

            elif action == 'deltaremantenimientolimpieza':
                try:
                    data['title'] = u'Eliminar Tarea Mantenimiento en Limpieza'
                    data['tareamantenimiento'] = MantenimientoGruCategoriaGarantiaLimp.objects.get(
                        pk=int(request.GET['id']))
                    return render(request, "inventario_activofijo/deltareamantenimientolimpieza.html", data)
                except:
                    pass

            elif action == 'edittaremantenimientolimpieza':
                try:
                    data['title'] = u'Editar tarea de mantenimiento'
                    data['tarea'] = tarea = MantenimientoGruCategoriaGarantiaLimp.objects.get(status=True, pk=int(
                        encrypt(request.GET['id'])))
                    data['form'] = TareaMantenimientoFrom(
                        initial={'categoria': tarea.grupocategoria, 'descripcion': tarea.descripcion})
                    return render(request, "inventario_activofijo/edittareamantenimientolimpieza.html", data)
                except:
                    pass

            elif action == 'tareasmantenimientodanio':
                try:
                    data['title'] = u'Lista de daños encontrados'
                    search = None
                    ids = None
                    if 's' in request.GET:
                        search = request.GET['s'].strip()
                        ss = search.split(' ')
                        if len(ss) == 1:
                            tareasdanio = MantenimientoGruCategoriaGarantiaErr.objects.filter(
                                descripcion__icontains=search, status=True).order_by('grupocategoria_id', 'descripcion')
                        else:
                            tareasdanio = MantenimientoGruCategoriaGarantiaErr.objects.filter(
                                descripcion__icontains=search, status=True).order_by('grupocategoria_id', 'descripcion')
                    else:
                        tareasdanio = MantenimientoGruCategoriaGarantiaErr.objects.filter(status=True).order_by(
                            'grupocategoria_id', 'descripcion')
                    paging = MiPaginador(tareasdanio, 15)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['tareasmantenimiento'] = page.object_list
                    return render(request, 'inventario_activofijo/view_tareasmantenimientosdanio.html', data)
                except Exception as ex:
                    pass

            elif action == 'addtareamantenimientodanio':
                try:
                    # puede_realizar_accion(request, 'sagest.puede_ingresar_compras')
                    data['title'] = u'Adicionar tarea de mantenimiento en daños'
                    data['form'] = TareaMantenimientoFrom()
                    plantilla = render(request, 'inventario_activofijo/addtareamantenimientodanio.html', data)
                    return plantilla
                except Exception as ex:
                    pass

            elif action == 'deltaremantenimientodanio':
                try:
                    data['title'] = u'Eliminar Tarea Mantenimiento en Daños'
                    data['tareamantenimiento'] = MantenimientoGruCategoriaGarantiaErr.objects.get(
                        pk=int(request.GET['id']))
                    return render(request, "inventario_activofijo/deltareamantenimientodanio.html", data)
                except:
                    pass

            elif action == 'edittaremantenimientodanio':
                try:
                    data['title'] = u'Editar tarea de mantenimiento'
                    data['tarea'] = tarea = MantenimientoGruCategoriaGarantiaErr.objects.get(status=True, pk=int(
                        encrypt(request.GET['id'])))
                    data['form'] = TareaMantenimientoFrom(
                        initial={'categoria': tarea.grupocategoria, 'descripcion': tarea.descripcion})
                    return render(request, "inventario_activofijo/edittareamantenimientodanio.html", data)
                except:
                    pass

            elif action == 'garantiamantenimiento':
                try:
                    data['title'] = u'Mantenimientos preventivos con garantía'
                    data['tipos'] = GruposCategoria.objects.all()
                    search = None
                    ids = None
                    garantiamantenimiento = None
                    if 's' in request.GET:
                        search = request.GET['s'].strip()
                        if search.isdigit():
                            garantiamantenimiento = MantenimientosActivosGarantia.objects.filter(
                                Q(activofijo__codigogobierno=search) | Q(activofijo__codigointerno=search) | Q(
                                    detallemantenimientosactivosgarantia__activofijo__codigogobierno=search) | Q(
                                    detallemantenimientosactivosgarantia__activofijo__codigointerno=search),
                                status=True)
                        else:
                            ss = search.split(' ')
                            if len(ss) == 1:
                                garantiamantenimiento = MantenimientosActivosGarantia.objects.filter(
                                    Q(activofijo__descripcion__icontains=search) | Q(
                                        detallemantenimientosactivosgarantia__activofijo__descripcion__icontains=search),
                                    status=True)
                            else:
                                garantiamantenimiento = MantenimientosActivosGarantia.objects.filter(
                                    Q(activofijo__descripcion__icontains=search) | Q(
                                        detallemantenimientosactivosgarantia__activofijo__descripcion__icontains=search),
                                    status=True)
                    else:
                        if 'id' in request.GET:
                            ids = request.GET['id']
                            garantiamantenimiento = MantenimientosActivosGarantia.objects.filter(pk=ids, status=True)
                        else:
                            if 'hist' in request.GET:
                                if request.GET['hist'] == '1':
                                    garantiamantenimiento = MantenimientosActivosGarantia.objects.filter(
                                        tipoactivo__descripcion='COMPUTADOR DE ESCRITORIO', status=True)
                                    data['histbus'] = '1'
                                else:
                                    garantiamantenimiento = MantenimientosActivosGarantia.objects.filter(status=True)
                                    data['histbus'] = '2'
                    if 'tipo' in request.GET:
                        tipo = request.GET['tipo']
                        data['idtipo'] = int(tipo)
                        data['histbus'] = '1'
                        garantiamantenimiento = MantenimientosActivosGarantia.objects.filter(tipoactivo__id=tipo,
                                                                                             status=True)
                    if ('fini' in request.GET) and ('ffin' in request.GET):
                        fini = request.GET['fini']
                        ffin = request.GET['ffin']
                        if 'hist' in request.GET:
                            hist = request.GET['hist']
                            if hist == '1':
                                tipo = request.GET['tipo']
                                garantiamantenimiento = MantenimientosActivosGarantia.objects.filter(
                                    tipoactivo__id=tipo, fechainicio__range=[fini, ffin], status=True)
                            elif hist == '2':
                                garantiamantenimiento = MantenimientosActivosGarantia.objects.filter(
                                    fechainicio__range=[fini, ffin], status=True)
                    paging = MiPaginador(garantiamantenimiento, 15)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['totalact'] = garantiamantenimiento.count()
                    data['garantiamantenimiento'] = page.object_list
                    return render(request, 'inventario_activofijo/garantiamantenimiento.html', data)
                except Exception as ex:
                    pass

            elif action == 'catalogo':
                try:
                    data['title'] = u'Catálogo Bienes'
                    search = None
                    ids = None
                    if 's' in request.GET:
                        search = request.GET['s'].strip()
                        ss = search.split(' ')
                        if len(ss) == 1:
                            catalogos = CatalogoBien.objects.filter(
                                Q(descripcion__icontains=search)).distinct().order_by('fechaingreso')
                        else:
                            catalogos = Edificio.objects.filter(Q(responsable__apellido1__icontains=ss[0]) | Q(
                                responsable__apellido2__icontains=ss[1]) |
                                                                (Q(identificacion__descripcion__icontains=ss[0]) & Q(
                                                                    identificacion__descripcion__icontains=ss[
                                                                        1]))).distinct().order_by('fechaingreso')
                    elif 'id' in request.GET:
                        ids = int(request.GET['id'])
                        catalogos = CatalogoBien.objects.filter(id=ids)
                    else:
                        catalogos = CatalogoBien.objects.filter(status=True).order_by('descripcion')
                    paging = MiPaginador(catalogos, 15)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['catalogos'] = page.object_list
                    return render(request, 'inventario_activofijo/view_catalogo.html', data)
                except Exception as ex:
                    pass

            elif action == 'inventariotecnologico':
                try:
                    data['title'] = u'Listado'
                    return render(request, 'inventario_activofijo/listadoinventario.html', data)
                except Exception as ex:
                    pass


            elif action == 'inventario':
                try:
                    search = None
                    ids = None
                    s = None
                    perfil = None
                    data['title'] = u'Listado de Activos Tecnológicos'
                    data['codigo'] = codigo = int(request.GET['codigo'])
                    data['rangosemaforo'] = anios = RangoVidaUtil.objects.filter(status=True).order_by('anio',
                                                                                                       'descripcion')
                    data['grupocatalogo'] = grupocatalogo = GruposCategoria.objects.filter(status=True)
                    filtrocatalogo = Q(status=True)
                    filtro = Q(status=True)
                    if not codigo == 0:
                        filtrocatalogo = filtrocatalogo & ((Q(archivobaja__isnull=True) |
                                                            Q(archivobaja='')) &
                                                           Q(catalogo__equipoelectronico=True) &
                                                           Q(catalogo__status=True) &
                                                           Q(catalogo__grupo__id=codigo) &
                                                           Q(statusactivo=1) &
                                                           Q(status=True))
                    else:
                        filtrocatalogo = filtrocatalogo & ((Q(archivobaja__isnull=True) |
                                                            Q(archivobaja='')) &
                                                           Q(catalogo__equipoelectronico=True) &
                                                           Q(catalogo__status=True) &
                                                           Q(statusactivo=1) &
                                                           Q(status=True))
                    # activo = ActivoFijo.objects.all()
                    if 's' in request.GET:
                        s = request.GET['s']
                        data['s'] = s

                        search = request.GET['s'].strip()
                        ss = search.split(' ')
                        if len(ss) == 1:
                            filtro = filtrocatalogo & ((Q(responsable__nombres__icontains=search) |
                                                        Q(responsable__apellido1__icontains=search) |
                                                        Q(responsable__apellido2__icontains=search) |
                                                        Q(responsable__cedula__icontains=search) |
                                                        Q(responsable__pasaporte__icontains=search)) |
                                                       Q(descripcion__icontains=search) |
                                                       Q(codigogobierno__icontains=search))
                        else:
                            filtro = filtrocatalogo & ((Q(responsable__apellido1__icontains=ss[0]) &
                                                        Q(responsable__apellido2__icontains=ss[1])) |
                                                       Q(descripcion__icontains=search) |
                                                       Q(codigogobierno__icontains=search))
                    else:
                        if 'cod' in request.GET:
                            filtro = filtro & (Q(pk=int(request.GET['cod'])))
                        else:
                            filtro = filtrocatalogo
                    activos = ActivoFijo.objects.filter(filtro).distinct().order_by('descripcion')
                    data['totales'] = totales = activos.values('id').count()
                    paging = MiPaginador(activos, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['perfil'] = perfil if perfil else ""
                    data['listadocatalogo'] = page.object_list
                    data['grupocatalogo'] = grupocatalogo
                    data['totales'] = totales
                    return render(request, "inventario_activofijo/view.html", data)
                except Exception as ex:
                    pass


            elif action == 'addresponsable':
                try:
                    form = ResponsableActivoForm()
                    data['responsable'] = responsable = int(request.GET['responsable'])
                    data['activo'] = activo = int(request.GET['activo'])
                    form.fields['persona'].queryset = Persona.objects.filter(Q(status=True) & (
                                Q(perfilusuario__administrativo__isnull=False) | Q(
                            perfilusuario__profesor__isnull=False))).distinct().order_by('apellido1', 'apellido2',
                                                                                         'nombres')
                    form.fields['responsableactual'].initial = Persona.objects.get(pk=responsable, status=True)
                    form.fields['activo'].initial = ActivoTecnologico.objects.get(pk=activo, status=True)
                    data['form2'] = form
                    data['action'] = 'addresponsable'
                    template = get_template('inventario_activofijo/modal/formasignarresponsable.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'addinventario':
                try:
                    form = InventarioActivoForm()
                    data['activo'] = activo = int(request.GET['activo'])
                    form.ocultarActivo()
                    form.fields['activo'].initial = ActivoFijo.objects.get(pk=activo, status=True)
                    data['form2'] = form
                    data['action'] = 'addinventario'
                    template = get_template('inventario_activofijo/modal/formregistroinventario.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'adicionarinventario':
                try:
                    form = InventarioActivoForm()
                    form.editaractivo()
                    form.fields['activotecnologico'].queryset = ActivoTecnologico.objects.filter(
                                                                                           Q(catalogo__equipoelectronico=True) &
                                                                                           Q(catalogo__status=True) &
                                                                                           Q(statusactivo=1) &
                                                                                           Q(status=True))
                    form.fields['estadogeneral'].queryset = EstadosGeneralesInventarioAT.objects.filter(status=True)
                    data['form2'] = form
                    data['action'] = 'adicionarinventario'
                    template = get_template('inventario_activofijo/modal/formregistroinventario.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass


            elif action == 'editarregistroinventario':
                try:
                    form = InventarioActivoForm()
                    inventario = ActivoFijoInventarioTecnologico.objects.get(status=True, pk=int(request.GET['activo']))
                    form.ocultarActivo()
                    form.fields['activo'].initial = ActivoTecnologico.objects.get(id=inventario.activotecnologico_id, status=True)
                    form.fields['observacion'].initial = inventario.observacion
                    form.fields['estadoubicacion'].initial = inventario.estadoubicacion
                    form.fields['estadofuncionamiento'].initial = inventario.estadofuncionamiento
                    form.fields['estadouso'].initial = inventario.estadouso
                    # form.fields['movimiento'].initial = inventario.movimiento
                    # form.fields['responsableactual'].initial = str(inventario.activo.responsable.__str__())
                    # inventarioestados = InventarioATEstadosGenerales.objects.filter(inventarioat_id=inventario.id)
                    # for inventarioestadogeneral in inventarioestados:
                    # form.fields['estadogeneral'].initial = list(InventarioATEstadosGenerales.objects.filter(inventarioat_id=inventario.id).values_list( 'id', flat=True))
                    form.cargarmultipleestadogeneral(inventario.id)
                    data['form2'] = form
                    data['activo'] = inventario.id
                    data['action'] = 'editarregistroinventario'
                    template = get_template('inventario_activofijo/modal/formregistroinventario.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass


            elif action == 'inventariotics':
                try:
                    search = None
                    ids = None
                    perfil = None
                    constatacion = 0
                    data['title'] = u'Inventario'
                    data['activos'] = activos = ActivoFijoInventarioTecnologico.objects.filter(status=True)
                    data['idperiodoinv'] = request.GET['idperiodoinv']
                    data['periodoinventarioat'] = periodoexistente = PeriodoInventarioAT.objects.filter(status=True, id=int(request.GET['idperiodoinv']))
                    if periodoexistente:
                        data['periodoinventario'] = periodoexistente[0].__str__()
                    filtrocatalogo = (Q(status=True) & Q(periodo_id=int(request.GET['idperiodoinv'])))
                    filtro = Q(status=True)
                    form = InventarioActivoForm()
                    form.editar()
                    form.editarobservacion()
                    form.fields['activo'].widget.attrs['disabled'] = False
                    data['form2'] = form
                    if 's' in request.GET:
                        s = request.GET['s']
                        data['s'] = s

                        search = request.GET['s'].strip()
                        ss = search.split(' ')
                        if len(ss) == 1:
                            filtro = filtrocatalogo & ((Q(activotecnologico__responsable__nombres__icontains=search) |
                                                        Q(activotecnologico__responsable__apellido1__icontains=search) |
                                                        Q(activotecnologico__responsable__apellido2__icontains=search) |
                                                        Q(activotecnologico__responsable__cedula__icontains=search) |
                                                        Q(activotecnologico__responsable__pasaporte__icontains=search)) |
                                                       Q(activotecnologico__descripcion__icontains=search) |
                                                       Q(activotecnologico__codigogobierno__icontains=search))
                        else:
                            filtro = filtrocatalogo & ((Q(activotecnologico__responsable__apellido1__icontains=ss[0]) &
                                                        Q(activotecnologico__responsable__apellido2__icontains=ss[1])) |
                                                       Q(activotecnologico__descripcion__icontains=search) |
                                                       Q(activotecnologico__codigogobierno__icontains=search))
                    else:
                        if 'cod' in request.GET:
                            filtro = filtro & (Q(pk=int(request.GET['cod'])))
                        else:
                            filtro = filtrocatalogo
                    activos = ActivoFijoInventarioTecnologico.objects.filter(filtro).distinct().order_by('-id')
                    if activos:
                        constatacion = activos.count()
                    data['totales'] = totales = activos.values('id').count()
                    paging = MiPaginador(activos, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['perfil'] = perfil if perfil else ""
                    data['listadocatalogo'] = page.object_list
                    data['totales'] = totales
                    data['constatacion'] = constatacion
                    return render(request, "inventario_activofijo/inventariotics.html", data)
                except Exception as ex:
                    pass

            elif action == 'cronograma':
                try:
                    search = None
                    ids = None
                    perfil = None
                    constatacion = 0
                    data['title'] = u'Inventario Tics'
                    data['activos'] = activos = ActivoFijoInventarioTecnologico.objects.filter(status=True)
                    # data['idperiodoinv'] = request.GET['idperiodoinv']
                    data['periodoinventarioat'] = periodoexistente = PeriodoInventarioAT.objects.filter(status=True, id=int(request.GET['idperiodoinv']))
                    if periodoexistente:
                        data['periodoinventario'] = periodoexistente[0].__str__()
                    filtrocatalogo = (Q(status=True) & Q(periodo_id=int(request.GET['idperiodoinv'])))
                    filtro = Q(status=True)
                    form = InventarioActivoForm()
                    form.editar()
                    form.editarobservacion()
                    form.fields['activo'].widget.attrs['disabled'] = False
                    data['form2'] = form
                    if 's' in request.GET:
                        s = request.GET['s']
                        data['s'] = s

                        search = request.GET['s'].strip()
                        ss = search.split(' ')
                        if len(ss) == 1:
                            filtro = filtrocatalogo & ((Q(activotecnologico__responsable__nombres__icontains=search) |
                                                        Q(activotecnologico__responsable__apellido1__icontains=search) |
                                                        Q(activotecnologico__responsable__apellido2__icontains=search) |
                                                        Q(activotecnologico__responsable__cedula__icontains=search) |
                                                        Q(activotecnologico__responsable__pasaporte__icontains=search)) |
                                                       Q(activotecnologico__descripcion__icontains=search) |
                                                       Q(activotecnologico__codigogobierno__icontains=search))
                        else:
                            filtro = filtrocatalogo & ((Q(activotecnologico__responsable__apellido1__icontains=ss[0]) &
                                                        Q(activotecnologico__responsable__apellido2__icontains=ss[1])) |
                                                       Q(activotecnologico__descripcion__icontains=search) |
                                                       Q(activotecnologico__codigogobierno__icontains=search))
                    else:
                        if 'cod' in request.GET:
                            filtro = filtro & (Q(pk=int(request.GET['cod'])))
                        else:
                            filtro = filtrocatalogo
                    activos = ActivoFijoInventarioTecnologico.objects.filter(filtro).distinct().order_by('-id')
                    if activos:
                        constatacion = activos.count()
                    data['totales'] = totales = activos.values('id').count()
                    paging = MiPaginador(activos, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['perfil'] = perfil if perfil else ""
                    data['listadocatalogo'] = page.object_list
                    data['totales'] = totales
                    data['constatacion'] = constatacion
                    return render(request, "inventario_activofijo/inventariotics.html", data)
                except Exception as ex:
                    pass


            elif action == 'traspasoconfirmado':
                try:
                    data['title'] = u'Confirmar traspaso'
                    data['responsable'] = int(request.GET['responsable'])
                    data['traspaso'] = ActivoFijo.objects.get(pk=int(request.GET['id']))
                    return render(request, "inventario_activofijo/traspasoconfirmado.html", data)
                except:
                    pass


            elif action == 'traspasoeliminar':
                try:
                    data['title'] = u'Confirmar eliminación'
                    data['traspaso'] = int(request.GET['id'])
                    return render(request, "inventario_activofijo/traspasoeliminar.html", data)
                except:
                    pass


            elif action == 'historialtraspaso':
                try:
                    if 'id' in request.GET:
                        data['historialtraspaso'] = HistorialTraspaso.objects.filter(activofijo=int(request.GET['id']),
                                                                                     status=True).order_by('-id')
                        data['activofijo'] = ActivoFijo.objects.get(status=True, pk=int(request.GET['id']))
                        template = get_template("inventario_activofijo/historialtraspaso.html")
                        return JsonResponse({"result": 'ok', 'data': template.render(data)})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})


            elif action == 'solicitudactivo':
                try:
                    if 'id' in request.GET:
                        data['estado'] = False
                        data['activo'] = ActivoFijo.objects.filter(status=True, pk=int(request.GET['id']))
                        data['solicitudactivo'] = SolicitudActivos.objects.filter(activo_id=int(request.GET['id']),
                                                                                  estado=4,
                                                                                  status=True).order_by('-id')
                        template = get_template("inventario_activofijo/solicitudactivo.html")
                        return JsonResponse({"result": 'ok', 'data': template.render(data)})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})


            elif action == 'solicitudactivoresponsable':
                try:
                    if 'id' in request.GET:
                        data['activo'] = ActivoFijo.objects.filter(status=True, pk=int(request.GET['id']))
                        data['solicitudactivo'] = SolicitudActivos.objects.filter(activo_id=int(request.GET['id']),
                                                                                  estado=1,
                                                                                  status=True).order_by('-id')
                        template = get_template("inventario_activofijo/solicitudactivoresponsable.html")
                        return JsonResponse({"result": 'ok', 'data': template.render(data)})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})


            elif action == 'solicitudtraspaso':
                try:
                    if 'id' in request.GET:
                        data['estado'] = True
                        data['activo'] = ActivoFijo.objects.filter(status=True, pk=int(request.GET['id']))
                        data['solicitudactivo'] = SolicitudActivos.objects.filter(activo_id=int(request.GET['id']),
                                                                                  estado=2, status=True).order_by('-id')
                        template = get_template("inventario_activofijo/solicitudactivo.html")
                        return JsonResponse({"result": 'ok', 'data': template.render(data)})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})


            elif action == 'consultarresponsable':
                try:
                    if 'id' in request.GET:
                        activo = ActivoFijo.objects.filter(status=True, pk=int(request.GET['id']))
                        responsable = activo[0].responsable.nombres + ' ' + activo[0].responsable.apellido1 + ' ' + \
                                      activo[0].responsable.apellido2
                        return JsonResponse({"result": 'ok', 'responsable': responsable})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})


            elif action == 'reportegeneralinventario':
                try:
                    totalconstatacion = None
                    filterestado = Q(status=True)
                    filter = (Q(status=True) & Q(periodo_id=int(request.GET['idperiodoinv'])))
                    idubicacion = int(request.GET['idubicacion'])
                    idfuncionamiento = int(request.GET['idfuncionamiento'])
                    # idestadogeneral = eval(request.GET.get('idestadogeneral'))
                    if not idubicacion == 3:
                        filter = filter & Q(estadoubicacion=idubicacion)
                    if not idfuncionamiento == 4:
                        filter = filter & Q(estadofuncionamiento=idfuncionamiento)
                    resultadoestados = ActivoFijoInventarioTecnologico.objects.filter(filter).values('id')
                    # if not len(idestadogeneral) == 0:
                    #     filterestado = Q(status=True)
                    #     listadoestadosg = []
                    #     resultadoinventarioestados = None
                    #
                    #     filterestado = filterestado & Q(estadogeneral_id__in=idestadogeneral)
                    #     for activoestado in resultadoestados:
                    #         resultadoinventarioestados = InventarioATEstadosGenerales.objects.filter(Q(inventarioat_id=activoestado['id']), Q(filterestado))
                    #         if (resultadoinventarioestados.count() == len(idestadogeneral)):
                    #             listadoestadosg.append(activoestado['id'])

                        # filtroporestado = None
                        # listadoestados = []
                        # filterestado = None
                        # for estadoat in idestadogeneral:
                        #     contador = 0
                        #     estadosgeneralesat = InventarioATEstadosGenerales.objects.filter(estadogeneral_id=estadoat).values('inventarioat_id')
                        #     # if estadosgeneralesat:
                        #     #     filterestado = filterestado & Q(id__in=estadosgeneralesat)
                        #     # if (len(estadosgeneralesat) == len(idestadogeneral)):
                        #     #     filterestado = filterestado & Q(id__in=estadosgeneralesat)
                        # # estadosgeneralesat = InventarioATEstadosGenerales.objects.filter(filterestado).values('inventarioat_id')
                        # filter = filter & Q(id__in=listadoestadosg)
                    data['listadoactivos'] = activos = ActivoFijoInventarioTecnologico.objects.filter(filter).order_by('-id')
                    totalactivos = ActivoTecnologico.objects.filter(Q(activotecnologico__archivobaja__isnull=True) |
                                                                Q(activotecnologico__archivobaja=''),
                                                                activotecnologico__catalogo__equipoelectronico=True,
                                                                activotecnologico__catalogo__status=True,
                                                                activotecnologico__statusactivo=1,
                                                                activotecnologico__status=True).count()
                    fechaminima = ActivoFijoInventarioTecnologico.objects.filter(status=True).aggregate(lowest=Min('fechainventario'))
                    data['fechaminima'] = fechaminima['lowest']
                    fechamaxima = ActivoFijoInventarioTecnologico.objects.filter(status=True).aggregate(maximo=Max('fechainventario'))
                    data['fechamaxima'] = fechamaxima['maximo']
                    totalconstatacion = activos.count()
                    if not totalactivos:
                        totalactivos = 0
                    if not activos:
                        totalconstatacion = 0
                    data['totalactivos'] = totalactivos
                    data['usuariogenera'] = usuario.username
                    data['totalconstatacion'] = totalconstatacion
                    data['fecha'] = datetime.now().date()
                    return conviert_html_to_pdf('inventario_activofijo/reporteinventarioactivo.html',
                                                {'pagesize': 'A4', 'data': data})
                except Exception as ex:
                    pass


            elif action == 'solicitudestraspasoresponsable':
                try:
                    search = None
                    ids = None
                    s = None
                    perfil = None
                    data['title'] = u'Pendientes'
                    filtrocatalogo = Q(status=True)
                    if 's' in request.GET:
                        s = request.GET['s']
                        data['s'] = s
                        search = request.GET['s'].strip()
                        ss = search.split(' ')
                        if len(ss) == 1:
                            filtro = filtrocatalogo & ((Q(responsableasignacion__nombres__icontains=search) |
                                                        Q(responsableasignacion__apellido1__icontains=search) |
                                                        Q(responsableasignacion__apellido2__icontains=search) |
                                                        Q(responsableasignacion__cedula__icontains=search) |
                                                        Q(responsableasignacion__pasaporte__icontains=search)) |
                                                       Q(activo__descripcion__icontains=search) |
                                                       Q(activo__codigogobierno__icontains=search))
                        else:
                            filtro = filtrocatalogo & ((Q(responsableasignacion__apellido1__icontains=ss[0]) &
                                                        Q(responsableasignacion__apellido2__icontains=ss[1])) |
                                                       Q(activo__descripcion__icontains=search) |
                                                       Q(activo__codigogobierno__icontains=search))
                    else:
                        filtro = filtrocatalogo
                    activos = SolicitudActivos.objects.filter(filtro).distinct().order_by('-id')
                    data['totales'] = totales = activos.values('id').count()
                    paging = MiPaginador(activos, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['perfil'] = perfil if perfil else ""
                    data['listadocatalogo'] = page.object_list
                    data['totales'] = totales
                    return render(request, "inventario_activofijo/solicitudestraspasoresponsable.html", data)
                except Exception as ex:
                    pass


            elif action == 'solicitudespendientesactivos':
                try:
                    search = None
                    ids = None
                    s = None
                    perfil = None
                    data['title'] = u'Pendientes'
                    filtrocatalogo = Q(status=True, estado=4)
                    if 's' in request.GET:
                        s = request.GET['s']
                        data['s'] = s
                        search = request.GET['s'].strip()
                        ss = search.split(' ')
                        if len(ss) == 1:
                            filtro = filtrocatalogo & ((Q(responsableasignacion__nombres__icontains=search) |
                                                        Q(responsableasignacion__apellido1__icontains=search) |
                                                        Q(responsableasignacion__apellido2__icontains=search) |
                                                        Q(responsableasignacion__cedula__icontains=search) |
                                                        Q(responsableasignacion__pasaporte__icontains=search)) |
                                                       Q(activo__descripcion__icontains=search) |
                                                       Q(activo__codigogobierno__icontains=search))
                        else:
                            filtro = filtrocatalogo & ((Q(responsableasignacion__apellido1__icontains=ss[0]) &
                                                        Q(responsableasignacion__apellido2__icontains=ss[1])) |
                                                       Q(activo__descripcion__icontains=search) |
                                                       Q(activo__codigogobierno__icontains=search))
                    else:
                        filtro = filtrocatalogo
                    activos = SolicitudActivos.objects.filter(filtro).distinct().order_by('-id')
                    data['totales'] = totales = activos.values('id').count()
                    paging = MiPaginador(activos, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['perfil'] = perfil if perfil else ""
                    data['listadocatalogo'] = page.object_list
                    data['totales'] = totales
                    return render(request, "inventario_activofijo/solicitudesactivos.html", data)
                except Exception as ex:
                    pass

            elif action == 'solicitudespendientestraspasos':
                try:
                    search = None
                    ids = None
                    s = None
                    perfil = None
                    data['title'] = u'Traspasos Pendientes'
                    filtrocatalogo = Q(status=True, estado=2)
                    if 's' in request.GET:
                        s = request.GET['s']
                        data['s'] = s
                        search = request.GET['s'].strip()
                        ss = search.split(' ')
                        if len(ss) == 1:
                            filtro = filtrocatalogo & ((Q(responsableasignacion__nombres__icontains=search) |
                                                        Q(responsableasignacion__apellido1__icontains=search) |
                                                        Q(responsableasignacion__apellido2__icontains=search) |
                                                        Q(responsableasignacion__cedula__icontains=search) |
                                                        Q(responsableasignacion__pasaporte__icontains=search)) |
                                                       Q(activo__descripcion__icontains=search) |
                                                       Q(activo__codigogobierno__icontains=search))
                        else:
                            filtro = filtrocatalogo & ((Q(responsableasignacion__apellido1__icontains=ss[0]) &
                                                        Q(responsableasignacion__apellido2__icontains=ss[1])) |
                                                       Q(activo__descripcion__icontains=search) |
                                                       Q(activo__codigogobierno__icontains=search))
                    else:
                        filtro = filtrocatalogo
                    activos = SolicitudActivos.objects.filter(filtro).distinct().order_by('-id')
                    data['totales'] = totales = activos.values('id').count()
                    paging = MiPaginador(activos, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['perfil'] = perfil if perfil else ""
                    data['listadocatalogo'] = page.object_list
                    data['totales'] = totales
                    return render(request, "inventario_activofijo/solicitudestraspasosactivos.html", data)
                except Exception as ex:
                    pass

            elif action == 'solicitudesconfirmadostraspasos':
                try:
                    search = None
                    ids = None
                    s = None
                    perfil = None
                    data['title'] = u'Traspasos Confirmados'
                    data['estado'] = True
                    filtrocatalogo = Q(status=True, estado=7)
                    if 's' in request.GET:
                        s = request.GET['s']
                        data['s'] = s
                        search = request.GET['s'].strip()
                        ss = search.split(' ')
                        if len(ss) == 1:
                            filtro = filtrocatalogo & ((Q(responsableasignacion__nombres__icontains=search) |
                                                        Q(responsableasignacion__apellido1__icontains=search) |
                                                        Q(responsableasignacion__apellido2__icontains=search) |
                                                        Q(responsableasignacion__cedula__icontains=search) |
                                                        Q(responsableasignacion__pasaporte__icontains=search)) |
                                                       Q(activo__descripcion__icontains=search) |
                                                       Q(activo__codigogobierno__icontains=search))
                        else:
                            filtro = filtrocatalogo & ((Q(responsableasignacion__apellido1__icontains=ss[0]) &
                                                        Q(responsableasignacion__apellido2__icontains=ss[1])) |
                                                       Q(activo__descripcion__icontains=search) |
                                                       Q(activo__codigogobierno__icontains=search))
                    else:
                        filtro = filtrocatalogo
                    activos = SolicitudActivos.objects.filter(filtro).distinct().order_by('-id')
                    data['totales'] = totales = activos.values('id').count()
                    paging = MiPaginador(activos, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['perfil'] = perfil if perfil else ""
                    data['listadocatalogo'] = page.object_list
                    data['totales'] = totales
                    return render(request, "inventario_activofijo/solicitudestraspasosactivos.html", data)
                except Exception as ex:
                    pass


            elif action == 'configuracioninventario':
                try:
                    search = None
                    ids = None
                    s = None
                    perfil = None
                    data['title'] = u'Configuración inventario'
                    return render(request, "inventario_activofijo/configuracioninventario.html", data)
                except Exception as ex:
                    pass


            elif action == 'periodoinventario':
                try:
                    search = None
                    ids = None
                    s = None
                    perfil = None
                    data['title'] = u'Periodos de inventario'
                    filtrocatalogo = Q(status=True)
                    if 's' in request.GET:
                        s = request.GET['s']
                        data['s'] = s
                        search = request.GET['s'].strip()
                        ss = search.split(' ')
                        if len(ss) == 1:
                            filtro = filtrocatalogo & ((Q(nombre__icontains=search) |
                                                        Q(detalle__icontains=search)))
                        else:
                            filtro = filtrocatalogo & ((Q(nombre__icontains=ss[0]) &
                                                        Q(nombre__icontains=ss[1])) |
                                                       (Q(detalle__icontains=ss[0]) &
                                                        Q(detalle__icontains=ss[1])))
                    else:
                        filtro = filtrocatalogo
                    periodos = PeriodoInventarioAT.objects.filter(filtro).distinct().order_by('-id')
                    data['totales'] = totales = periodos.values('id').count()
                    paging = MiPaginador(periodos, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['perfil'] = perfil if perfil else ""
                    data['listadocatalogo'] = page.object_list
                    data['totales'] = totales
                    return render(request, "inventario_activofijo/periodoinventario.html", data)
                except Exception as ex:
                    pass

            elif action == 'asignacioncierreinventario':
                try:
                    search = None
                    ids = None
                    s = None
                    perfil = None
                    data['title'] = u'Cierres de inventario'
                    data['idperiodoinv'] = int(request.GET['idperiodoinv'])
                    data['periodoinventarioat'] = periodoexistente = PeriodoInventarioAT.objects.get(status=True,id=int(request.GET['idperiodoinv']))
                    filtrocatalogo = Q(status=True)
                    if 's' in request.GET:
                        s = request.GET['s']
                        data['s'] = s
                        search = request.GET['s'].strip()
                        ss = search.split(' ')
                        if len(ss) == 1:
                            filtro = filtrocatalogo & ((Q(persona__nombres__icontains=search) |
                                                        Q(persona__apellido1__icontains=search) |
                                                        Q(persona__apellido2__icontains=search) |
                                                        Q(periodo__nombre__icontains=search)))
                        else:
                            filtro = filtrocatalogo & ((Q(persona__nombres__icontains=ss[0]) &
                                                        Q(persona__nombres__icontains=ss[1])) |
                                                       (Q(persona__apellido1__icontains=ss[0]) &
                                                        Q(persona__apellido2__icontains=ss[1])) |
                                                       (Q(periodo__nombre__icontains=ss[0]) &
                                                        Q(periodo__nombre__icontains=ss[1])))
                    else:
                        filtro = filtrocatalogo
                    periodos = AsignacionCierreInventarioAT.objects.filter(filtro).distinct().order_by('-id')
                    data['totales'] = totales = periodos.values('id').count()
                    paging = MiPaginador(periodos, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['perfil'] = perfil if perfil else ""
                    data['listadocatalogo'] = page.object_list
                    data['totales'] = totales
                    return render(request, "inventario_activofijo/asignacioncierreinventario.html", data)
                except Exception as ex:
                    pass

            elif action == 'addperiodoinventario':
                try:
                    form = PeriodoInventarioATForm()
                    data['form2'] = form
                    data['action'] = 'addperiodoinventario'
                    template = get_template('inventario_activofijo/modal/formregistroinventario.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'editperiodoinventario':
                try:
                    form = PeriodoInventarioATForm()
                    periodo = PeriodoInventarioAT.objects.get(status=True, pk=int(request.GET['periodo']))
                    form.fields['nombre'].initial = periodo.nombre
                    form.fields['fechainicio'].initial = periodo.fechainicio.strftime("%Y-%m-%d")
                    form.fields['fechafin'].initial = periodo.fechafin.strftime("%Y-%m-%d")
                    form.fields['estado'].initial = periodo.estado
                    form.fields['detalle'].initial = periodo.detalle
                    data['form2'] = form
                    data['idperiodo'] = int(request.GET['idperiodo'])
                    data['action'] = 'editperiodoinventario'
                    template = get_template('inventario_activofijo/modal/formregistroinventario.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass


            elif action == 'eliminaperiodo':
                try:
                    data['title'] = u'Confirmar eliminación'
                    data['idperiodo'] = int(request.GET['id'])
                    return render(request, "inventario_activofijo/eliminaperiodo.html", data)
                except:
                    pass


            elif action == 'activoinventario':
                try:
                    form = InventarioActivoForm()
                    forms = form.fields['activo']
                    data['form2'] = forms
                    data['action'] = 'reporteactivoinventario'
                    template = get_template('inventario_activofijo/modal/formreporteinventario.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass


            elif action == 'chartinventariotecnologico':
                try:
                    data['title'] = u'Equipo Tecnológico'
                    data['codigo'] = codigo = int(request.GET['codigo'])
                    data['rangosemaforo'] = anios = RangoVidaUtil.objects.filter(status=True).order_by('anio',
                                                                                                       'descripcion')
                    data['grupocatalogo'] = GruposCategoria.objects.filter(status=True)
                    if codigo == 0:
                        data['listadocatalogo'] = activos = ActivoTecnologico.objects.filter(
                           activotecnologico__statusactivo=1, status=True).distinct().order_by('descripcion')
                    else:
                        data['listadocatalogo'] = activos = ActivoTecnologico.objects.filter(
                            activotecnologico__statusactivo=1,activotecnologico__catalogo__grupo__id=codigo, status=True).distinct().order_by(
                            'descripcion')
                    data['totales'] = activos.values('id').count()
                    return render(request, "inventario_activofijo/charinventariotecnologico.html", data)
                except Exception as ex:
                    pass

            elif action == 'addedificio':
                try:
                    # puede_realizar_accion(request, 'sagest.puede_ingresar_compras')
                    data['title'] = u'Adicionar Bien Inmueble de la Institución'
                    data['form'] = EdificioForm()
                    plantilla = render(request, 'inventario_activofijo/addedificio.html', data)
                    return plantilla
                except Exception as ex:
                    pass

            elif action == 'addtareamantenimiento':
                try:
                    # puede_realizar_accion(request, 'sagest.puede_ingresar_compras')
                    data['title'] = u'Adicionar tarea de mantenimiento'
                    data['form'] = TareaMantenimientoFrom()
                    plantilla = render(request, 'inventario_activofijo/addtareamantenimiento.html', data)
                    return plantilla
                except Exception as ex:
                    pass

            elif action == 'edittaremantenimiento':
                try:
                    data['title'] = u'Editar tarea de mantenimiento'
                    data['tarea'] = tarea = MantenimientoGruCategoria.objects.get(status=True,
                                                                                  pk=int(request.GET['id']))
                    data['form'] = TareaMantenimientoFrom(
                        initial={'categoria': tarea.grupocategoria, 'descripcion': tarea.descripcion})
                    return render(request, "inventario_activofijo/edittareamantenimiento.html", data)
                except:
                    pass

            elif action == 'addmantenimientogarantia':
                try:
                    # puede_realizar_accion(request, 'sagest.puede_ingresar_compras')
                    data['title'] = u'Nuevo Mantenimiento Equipo Tecnológico Con Garantía'
                    data['form'] = MantenimientosActivosGarantiaForm()
                    data['form3'] = ActivosFijosForm()
                    data['grupocategoria'] = GruposCategoria.objects.filter(status=True)
                    plantilla = render(request, 'inventario_activofijo/addmantenimientogarantia.html', data)
                    return plantilla
                except Exception as ex:
                    pass

            elif action == 'editedificio':
                try:
                    data['title'] = u'Editar Bienes Inmuebles de la Institución'
                    data['edificio'] = edificio = Edificio.objects.get(pk=int(request.GET['id']))
                    initial = model_to_dict(edificio)
                    form = EdificioForm(initial=initial)
                    form.fields['catalogo'].widget.attrs['descripcion'] = edificio.catalogo
                    form.fields['catalogo'].widget.attrs['value'] = edificio.catalogo.id
                    form.fields['responsable'].widget.attrs['descripcion'] = edificio.responsable
                    form.fields['responsable'].widget.attrs['value'] = edificio.responsable.id
                    form.fields['custodio'].widget.attrs['descripcion'] = edificio.custodio
                    form.fields['custodio'].widget.attrs['value'] = edificio.custodio.id
                    data['form'] = form
                    return render(request, "inventario_activofijo/editedificio.html", data)
                except:
                    pass

            elif action == 'editmantenimientogarantia':
                try:
                    ingnuevo = request.GET['nuevo']
                    data['title'] = u'Editar Mantenimiento Garantía'
                    data['mangarantia'] = mangarantia = MantenimientosActivosGarantia.objects.get(
                        pk=int(encrypt(request.GET['idgarantia'])))
                    #
                    data['grupocategoria'] = GruposCategoria.objects.filter(status=True)
                    # Llamando a los componentes de daños y limpieza
                    data['tareaslimpieza'] = MantenimientoGruCategoriaGarantiaLimp.objects.filter(
                        grupocategoria=mangarantia.tipoactivo, status=True)
                    data['tareasdanio'] = MantenimientoGruCategoriaGarantiaErr.objects.filter(
                        grupocategoria=mangarantia.tipoactivo, status=True)
                    data['tareaslimpiezat'] = TareasActivosPreventivosGarantiaLimp.objects.values_list('grupos_id',
                                                                                                       flat=True).filter(
                        mantenimiento=mangarantia, status=True)
                    data['tareasdaniot'] = TareasActivosPreventivosGarantiaErr.objects.values_list('grupos_id',
                                                                                                   flat=True).filter(
                        mantenimiento=mangarantia, status=True)
                    #
                    data['detallemantenimiento'] = mangarantia.detallemantenimientosactivosgarantia_set.filter(
                        status=True).order_by('activofijo__descripcion')
                    form = MantenimientosActivosGarantiaForm(initial={'numreporte': mangarantia.numreporte,
                                                                      'fechainicio': mangarantia.fechainicio,
                                                                      'valor': mangarantia.valor,
                                                                      'tipoactivo': mangarantia.tipoactivo,
                                                                      'observacion': mangarantia.observacion,
                                                                      'horamax': mangarantia.horamax,
                                                                      'minutomax': mangarantia.minutomax,
                                                                      'estfrec': mangarantia.estfrec,
                                                                      'estfent': mangarantia.estfent})
                    data['form'] = form
                    # from sga.forms import ReservasCraiSolicitarAutoridadForm
                    data['form3'] = ActivosFijosForm()
                    if (ingnuevo == 'False'):
                        return render(request, "inventario_activofijo/editmantenimientogarantia.html", data)
                    else:
                        return render(request, "inventario_activofijo/editmantenimientogarantianuevo.html", data)
                except Exception as ex:
                    pass

            elif action == 'deledificio':
                try:
                    data['title'] = u'Eliminar Bienes Inmuebles de la Institución'
                    data['edificio'] = Edificio.objects.get(pk=int(request.GET['id']))
                    return render(request, "inventario_activofijo/deledificio.html", data)
                except:
                    pass

            elif action == 'delmantenimiento':
                try:
                    data['title'] = u'Eliminar Mantenimiento'
                    data['mantenimiento'] = MantenimientosActivosPreventivos.objects.get(pk=int(request.GET['id']))
                    return render(request, "inventario_activofijo/delmantenimiento.html", data)
                except:
                    pass

            elif action == 'delmantenimientogarantia':
                try:
                    data['title'] = u'Eliminar Mantenimiento Garantía'
                    data['mantenimientogarantia'] = MantenimientosActivosGarantia.objects.get(pk=int(request.GET['id']))
                    return render(request, "inventario_activofijo/delmantenimientogarantia.html", data)
                except:
                    pass

            elif action == 'deltaremantenimiento':
                try:
                    data['title'] = u'Eliminar Tarea Mantenimiento'
                    data['tareamantenimiento'] = MantenimientoGruCategoria.objects.get(pk=int(request.GET['id']))
                    return render(request, "inventario_activofijo/deltareamantenimiento.html", data)
                except:
                    pass

            elif action == 'buscarresponsable_custodio':
                try:
                    q = request.GET['q'].upper().strip()
                    s = q.split(" ")
                    if s.__len__() == 2:
                        persona = Persona.objects.filter(status=True, administrativo__isnull=False,
                                                         nombres__contains=s[0], apellido1__icontains=s[0],
                                                         apellido2__icontains=s[0]).distinct()[:20]
                    else:
                        persona = Persona.objects.filter((Q(nombres__contains=s[0]) | Q(apellido1__contains=s[0]) | Q(
                            apellido2__contains=s[0]) | Q(cedula__contains=s[0]))).filter(status=True,
                                                                                          administrativo__isnull=False).distinct()[
                                  :20]
                    data = {"result": "ok", "results": [{"id": x.id, "name": x.flexbox_repr()} for x in persona]}
                    return JsonResponse((data))
                except Exception as ex:
                    pass

            elif action == 'add':
                try:
                    # puede_realizar_accion(request, 'sagest.puede_ingresar_compras')
                    data['title'] = u'Nuevo Activo'
                    form = ActivoFijoForm()
                    data['form'] = form
                    plantilla = render(request, 'inventario_activofijo/add.html', data)
                    return plantilla
                except Exception as ex:
                    pass

            elif action == 'addmantenimiento':
                try:
                    # puede_realizar_accion(request, 'sagest.puede_ingresar_compras')
                    data['title'] = u'Nuevo Mantenimiento Equipo Tecnológico Sin Garantía'
                    # data['tiposmantenimiento'] = MantenimientoGruCategoria.objects.filter(status=True)
                    data['form'] = MantenimientosActivosPreventivosForm()
                    data['grupocategoria'] = GruposCategoria.objects.filter(status=True)
                    data['form3'] = ActivosFijosForm()
                    # data['tipo'] = TIPO_MANTENIMIENTO
                    return render(request, "inventario_activofijo/addmantenimiento.html", data)
                except Exception as ex:
                    pass

            elif action == 'editmantenimiento':
                try:
                    # puede_realizar_accion(request, 'sagest.puede_ingresar_compras')
                    data['title'] = u'Editar Mantenimiento'
                    data['manpreventivos'] = mantenimiento = MantenimientosActivosPreventivos.objects.get(
                        pk=request.GET['idman'])
                    data['grupocategoria'] = GruposCategoria.objects.filter(status=True)
                    data['tareasmantenimiento'] = MantenimientoGruCategoria.objects.filter(
                        grupocategoria=mantenimiento.tipoactivo, status=True, activo=True)
                    data['tareasactivo'] = TareasActivosPreventivos.objects.values_list('grupos_id', flat=True).filter(
                        mantenimiento=mantenimiento, status=True)
                    data['piezaparteactivo'] = PiezaParteActivosPreventivos.objects.filter(mantenimiento=mantenimiento,
                                                                                           status=True)
                    data['daniomantenimiento'] = MantenimientoGruDanios.objects.filter(
                        grupocategoria=mantenimiento.tipoactivo, status=True, activo=True)
                    data['danioactivo'] = TareasActivosPreventivosDanios.objects.filter(
                        mantenimiento=mantenimiento, status=True)
                    data['tipo'] = TIPO_MANTENIMIENTO
                    form = MantenimientosActivosPreventivosForm(initial={'tipoactivo': mantenimiento.tipoactivo,
                                                                         'estusu': mantenimiento.estusu,
                                                                         'archivo': mantenimiento.archivo,
                                                                         'fecha': mantenimiento.fecha,
                                                                         'horamax': mantenimiento.horamax,
                                                                         'minutomax': mantenimiento.minutomax,
                                                                         'estfrec': mantenimiento.funcionarecibe,
                                                                         'estfent': mantenimiento.funcionaentrega,
                                                                         'marca': mantenimiento.marca,
                                                                         'modelo': mantenimiento.modelo,
                                                                         # 'caracteristica': mantenimiento.caracteristicas,
                                                                         'bsugiere': mantenimiento.sbequipo,
                                                                         'dsugiere': mantenimiento.descbaja,
                                                                         # 'piezaparte': piezaparte,
                                                                         'observacion': mantenimiento.observaciones})
                    form.cargar_mantenimiento(mantenimiento)
                    # cat = HdPiezaPartes.objects.filter(pk=mantenimiento.tipoactivo)
                    # form.fields['piezaparte'].queryset = cat
                    data['form'] = form
                    data['form3'] = ActivosFijosForm()
                    if mantenimiento.nuevo:
                        return render(request, "inventario_activofijo/editmantenimientov2.html", data)
                    else:
                        return render(request, "inventario_activofijo/editmantenimiento.html", data)
                except Exception as ex:
                    pass

            elif action == 'edit':
                try:
                    data['title'] = u'Modificar Activo'
                    data['activo'] = activo = ActivoFijo.objects.get(pk=request.GET['id'])
                    form = ActivoFijoForm(initial={
                        'codigogobierno': activo.codigogobierno,
                        'codigointerno': activo.codigointerno,
                        'estructuraactivo': activo.estructuraactivo,
                        'clasebien': activo.clasebien,
                        'fechaingreso': activo.fechaingreso,
                        'descripcion': activo.descripcion,
                        'observacion': activo.observacion,
                        'origeningreso': activo.origeningreso,
                        'costo': activo.costo,
                        'serie': activo.serie,
                        'modelo': activo.modelo,
                        'marca': activo.marca,
                        'tipodocumentorespaldo': activo.tipodocumentorespaldo,
                        'clasedocumentorespaldo': activo.clasedocumentorespaldo,
                        'numerocomprobante': activo.numerocomprobante,
                        'tipocomprobante': activo.tipocomprobante,
                        'fechacomprobante': activo.fechacomprobante,
                        'estado': activo.estado,
                        'tipoproyecto': activo.tipoproyecto,
                        'deprecia': activo.deprecia,
                        'vidautil': activo.vidautil,
                        'cuentacontable': activo.cuentacontable,
                        'ubicacion': activo.ubicacion,
                        'titulo': activo.titulo,
                        'autor': activo.autor,
                        'editorial': activo.editorial,
                        'fechaedicion': activo.fechaedicion,
                        'numeroedicion': activo.numeroedicion,
                        'clasificacionbibliografica': activo.clasificacionbibliografica,
                        'color': activo.color,
                        'material': activo.material,
                        'dimensiones': activo.dimensiones,
                        'clasevehiculo': activo.clasevehiculo,
                        'tipovehiculo': activo.tipovehiculo,
                        'numeromotor': activo.numeromotor,
                        'numerochasis': activo.numerochasis,
                        'placa': activo.placa,
                        'aniofabricacion': activo.aniofabricacion,
                        'colorprimario': activo.colorprimario,
                        'colorsecundario': activo.colorsecundario,
                        'propietario': activo.propietario,
                        'codigocatastral': activo.codigocatastral,
                        'numeropredio': activo.numeropredio,
                        'valoravaluo': activo.valoravaluo,
                        'anioavaluo': activo.anioavaluo,
                        'areapredio': activo.areapredio,
                        'areaconstruccion': activo.areaconstruccion,
                        'pisos': activo.pisos,
                        'provincia': activo.provincia,
                        'canton': activo.canton,
                        'parroquia': activo.parroquia,
                        'zona': activo.zona,
                        'nomenclatura': activo.nomenclatura,
                        'sector': activo.sector,
                        'direccion': activo.direccion,
                        'direccion2': activo.direccion2,
                        'escritura': activo.escritura,
                        'fechaescritura': activo.fechaescritura,
                        'notaria': activo.notaria,
                        'beneficiariocontrato': activo.beneficiariocontrato,
                        'fechacontrato': activo.fechacontrato,
                        'duracioncontrato': activo.duracioncontrato,
                        'montocontrato': activo.montocontrato,
                        'catalogo': activo.catalogo.id,
                    })
                    form.editar(activo, request.user)
                    data['form'] = form
                    return render(request, 'inventario_activofijo/edit.html', data)
                except Exception as ex:
                    pass

            elif action == 'addingresoinformebaja':
                try:
                    data['title'] = u'Ingreso informe de baja'
                    data['activofijo'] = activofijo = ActivoFijo.objects.get(pk=int(encrypt(request.GET['idactivo'])))
                    data['tipo'] = tipo = (request.GET['tipo'])
                    data['formdetalle'] = DetalleInformeBajaForm()
                    if activofijo.responsable:
                        idsolicita = activofijo.responsable.id
                    else:
                        idsolicita = 0
                    data['codsolicita'] = idsolicita
                    if tipo == 1:
                        form = InformeBajaForm(initial={'solicita': activofijo.responsable,
                                                        'estadouso': 2,
                                                        'estado': 2})
                    else:
                        form = InformeBajaForm2(initial={'solicita': activofijo.responsable,
                                                         'estadouso': 2,
                                                         'estado': 2})
                    data['form'] = form
                    return render(request, 'inventario_activofijo/addingresobaja.html', data)
                except Exception as ex:
                    pass

            elif action == 'editingresoinformebaja':
                try:
                    data['title'] = u'Modificar Activo'
                    data['activoinformebaja'] = activoinformebaja = InformeActivoBaja.objects.get(
                        activofijo_id=int(encrypt(request.GET['idactivo'])))
                    if activoinformebaja.tipoinforme == 1:
                        form = InformeBajaForm(initial={
                            'solicita': activoinformebaja.solicita,
                            'responsable': activoinformebaja.responsable,
                            'conclusion': activoinformebaja.conclusion,
                            'estado': activoinformebaja.estado,
                            'estadouso': activoinformebaja.estadouso,
                            'bloque': activoinformebaja.bloque,
                            'detallerevision': activoinformebaja.detallerevision})
                    else:
                        form = InformeBajaForm2(initial={
                            'solicita': activoinformebaja.solicita,
                            'responsable': activoinformebaja.responsable,
                            'conclusion': activoinformebaja.conclusion,
                            'estado': activoinformebaja.estado,
                            'estadouso': activoinformebaja.estadouso,
                            'bloque': activoinformebaja.bloque,
                            'detallerevision': activoinformebaja.detallerevision,
                            'departamento': activoinformebaja.departamento,
                            'gestion': activoinformebaja.gestion})

                    data['form'] = form
                    data['formdetalle'] = DetalleInformeBajaForm()
                    data['detalleinformebaja'] = activoinformebaja.detalleinformeactivobaja_set.filter(status=True)
                    return render(request, 'inventario_activofijo/editingresobaja.html', data)
                except Exception as ex:
                    pass

            elif action == 'excellistadoactivos':
                try:
                    fdesde, fhasta = request.GET.get('fechadesde', ''), request.GET.get('fechahasta', '')
                    __author__ = 'Unemi'
                    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on',
                                      num_format_str='#,##0.00')
                    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                    title = easyxf(
                        'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    style1 = easyxf(num_format_str='D-MMM-YY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('exp_xls_post_part')
                    # ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    response = HttpResponse(content_type="application/ms-excel")
                    response[
                        'Content-Disposition'] = 'attachment; filename=Listado' + random.randint(1,
                                                                                                 10000).__str__() + '.xls'
                    columns = [
                        (u"RESPONSABLE", 10000),
                        (u"CODIGO INTERNO", 2000),
                        (u"CPDOGP GOBIERNO", 2000),
                        (u"FECHA INGRESO", 4000),
                        (u"TIEMPO", 4500),
                        (u"NUM. AÑOS", 2000),
                        (u"CATALOGO", 15000),
                        (u"DESCRIPCIÓN", 10000),
                        (u"MODELO", 10000),
                        (u"MARCA", 10000),
                        (u"DOCUMENTO BAJA", 10000),
                        (u"FECHA BAJA", 10000),
                        (u"UBICACIÓN", 10000),
                        (u"ESTADO", 10000),
                        (u"ARCHIVO DE BAJA", 10000),
                    ]
                    row_num = 0
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    date_format = xlwt.XFStyle()
                    date_format.num_format_str = 'yyyy/mm/dd'
                    lista1 = '0'
                    if 'cadenatexto' in request.GET:
                        lista1 = request.GET['cadenatexto']
                    if lista1 == '0' and not fdesde and not fhasta:
                        sql = "select ac.codigointerno,ac.codigogobierno,ac.fechaingreso,age (current_date, ac.fechaingreso) || '' tiempo, " \
                              "cast(extract(year from age (current_date, ac.fechaingreso))*12 + extract(month from age (ac.fechaingreso)) as int)/12 as numanios " \
                              ",cat.descripcion as catalogo,ac.descripcion,ac.modelo,ac.marca," \
                              "(select perso.apellido1 || ' ' || perso.apellido2 || ' ' || perso.nombres from sga_persona perso where id=ac.responsable_id) as responsable , " \
                              "ac.id, (SELECT ubi.nombre AS ubicacion FROM sagest_ubicacion ubi WHERE ubi.id = ac.ubicacion_id)," \
                              "(SELECT est.nombre AS estado FROM sagest_estadoproducto est WHERE est.id = ac.estado_id), ac.archivobaja " \
                              "from sagest_activofijo ac,sagest_catalogobien cat,sagest_gruposcategoria gru " \
                              "where ac.catalogo_id=cat.id and ac.statusactivo=1 and cat.clasificado=True and cat.equipoelectronico=True and cat.grupo_id=gru.id " \
                              " and ac.status=True and (ac.archivobaja='' or ac.archivobaja isnull)"
                    if fdesde and fhasta and lista1 == '0':
                        sql = "select ac.codigointerno,ac.codigogobierno,ac.fechaingreso,age (current_date, ac.fechaingreso) || '' tiempo, " \
                              "cast(extract(year from age (current_date, ac.fechaingreso))*12 + extract(month from age (ac.fechaingreso)) as int)/12 as numanios " \
                              ",cat.descripcion as catalogo,ac.descripcion,ac.modelo,ac.marca," \
                              "(select perso.apellido1 || ' ' || perso.apellido2 || ' ' || perso.nombres from sga_persona perso where id=ac.responsable_id) as responsable , " \
                              "ac.id, (SELECT ubi.nombre AS ubicacion FROM sagest_ubicacion ubi WHERE ubi.id = ac.ubicacion_id)," \
                              "(SELECT est.nombre AS estado FROM sagest_estadoproducto est WHERE est.id = ac.estado_id), ac.archivobaja " \
                              "from sagest_activofijo ac,sagest_catalogobien cat,sagest_gruposcategoria gru " \
                              "where ac.catalogo_id=cat.id and ac.statusactivo=1 and cat.clasificado=True and cat.equipoelectronico=True and cat.grupo_id=gru.id " \
                              " and ac.status=True and (ac.archivobaja='' or ac.archivobaja isnull) AND ac.fechaingreso >= '{}' AND ac.fechaingreso <= '{}'".format(
                            fdesde, fhasta)
                    if lista1 != '0' and not fdesde and not fhasta:
                        sql = "select ac.codigointerno,ac.codigogobierno,ac.fechaingreso,age (current_date, ac.fechaingreso) || '' tiempo, " \
                              "cast(extract(year from age (current_date, ac.fechaingreso))*12 + extract(month from age (ac.fechaingreso)) as int)/12 as numanios " \
                              ",cat.descripcion as catalogo,ac.descripcion,ac.modelo,ac.marca,(select perso.apellido1 || ' ' || perso.apellido2 || ' ' || perso.nombres from sga_persona perso where id=ac.responsable_id) as responsable ," \
                              " ac.id, (SELECT ubi.nombre AS ubicacion FROM sagest_ubicacion ubi WHERE ubi.id = ac.ubicacion_id), " \
                              " (SELECT est.nombre AS estado FROM sagest_estadoproducto est WHERE est.id = ac.estado_id), ac.archivobaja " \
                              "from sagest_activofijo ac,sagest_catalogobien cat,sagest_gruposcategoria gru " \
                              "where ac.catalogo_id=cat.id and ac.statusactivo=1 and cat.clasificado=True and cat.equipoelectronico=True and cat.grupo_id=gru.id " \
                              " and ac.status=True and gru.id=" + lista1 + " and (ac.archivobaja='' or ac.archivobaja isnull)"
                    if lista1 != '0' and fdesde and fhasta:
                        sql = "select ac.codigointerno,ac.codigogobierno,ac.fechaingreso,age (current_date, ac.fechaingreso) || '' tiempo, " \
                              "cast(extract(year from age (current_date, ac.fechaingreso))*12 + extract(month from age (ac.fechaingreso)) as int)/12 as numanios " \
                              ",cat.descripcion as catalogo,ac.descripcion,ac.modelo,ac.marca,(select perso.apellido1 || ' ' || perso.apellido2 || ' ' || perso.nombres from sga_persona perso where id=ac.responsable_id) as responsable ," \
                              " ac.id, (SELECT ubi.nombre AS ubicacion FROM sagest_ubicacion ubi WHERE ubi.id = ac.ubicacion_id), " \
                              " (SELECT est.nombre AS estado FROM sagest_estadoproducto est WHERE est.id = ac.estado_id), ac.archivobaja " \
                              "from sagest_activofijo ac,sagest_catalogobien cat,sagest_gruposcategoria gru " \
                              "where ac.catalogo_id=cat.id and ac.statusactivo=1 and cat.clasificado=True  and cat.equipoelectronico=True and cat.grupo_id=gru.id " \
                              " and ac.status=True and gru.id=" + lista1 + " and (ac.archivobaja='' or ac.archivobaja isnull) AND ac.fechaingreso >= '{}' AND ac.fechaingreso <= '{}'".format(
                            fdesde, fhasta)
                    cursor = connection.cursor()
                    cursor.execute(sql)
                    results = cursor.fetchall()
                    row_num = 1
                    for r in results:
                        campo1 = r[0]
                        campo2 = r[1]
                        campo3 = r[2]
                        campo4 = r[3]
                        campo5 = r[4]
                        campo6 = r[5]
                        campo7 = r[6]
                        campo8 = r[7]
                        campo9 = r[8]
                        campo10 = r[9]
                        campo11 = 'NO'
                        campo12 = ''
                        if InformeActivoBaja.objects.filter(status=True, activofijo_id=r[10]).exists():
                            informe = InformeActivoBaja.objects.get(status=True, activofijo_id=r[10])
                            campo11 = 'SI'
                            campo12 = str(informe.fecha_creacion)
                        campo13 = r[11]
                        campo14 = r[12]
                        campo15 = r[13]
                        ws.write(row_num, 0, campo10, font_style2)
                        ws.write(row_num, 1, campo1, font_style2)
                        ws.write(row_num, 2, campo2, font_style2)
                        ws.write(row_num, 3, campo3, date_format)
                        ws.write(row_num, 4, campo4, font_style2)
                        ws.write(row_num, 5, campo5, font_style2)
                        ws.write(row_num, 6, campo6, font_style2)
                        ws.write(row_num, 7, campo7, font_style2)
                        ws.write(row_num, 8, campo8, font_style2)
                        ws.write(row_num, 9, campo9, font_style2)
                        ws.write(row_num, 10, campo11, font_style2)
                        ws.write(row_num, 11, campo12, font_style2)
                        ws.write(row_num, 12, campo13, font_style2)
                        ws.write(row_num, 13, campo14, font_style2)
                        ws.write(row_num, 14, "https://sagest.unemi.edu.ec/media/".format(campo15), font_style2)
                        row_num += 1
                    wb.save(response)
                    return response
                except Exception as ex:
                    pass

            elif action == 'pdflistadoactivos':
                try:
                    data['title'] = nombrearchivo = u'Informe de Activos Tecnológicos'
                    fdesde, fhasta = request.GET.get('fechadesde', ''), request.GET.get('fechahasta', '')
                    data['desde'] = str(fdesde)
                    data['hasta'] = str(fhasta)
                    lista1 = '0'
                    if 'cadenatexto' in request.GET:
                        lista1 = request.GET['cadenatexto']
                    values_ac = InformeActivoBaja.objects.values('activofijo__id').filter(status=True).values_list(
                        'activofijo__id', flat=True)
                    if lista1 == '0':
                        sql = "select ac.codigointerno,ac.codigogobierno,ac.fechaingreso,age (current_date, ac.fechaingreso) || '' tiempo, " \
                              "cast(extract(year from age (current_date, ac.fechaingreso))*12 + extract(month from age (ac.fechaingreso)) as int)/12 as numanios " \
                              ",cat.descripcion as catalogo,ac.descripcion,ac.modelo,ac.marca," \
                              "(select perso.apellido1 || ' ' || perso.apellido2 || ' ' || perso.nombres from sga_persona perso where id=ac.responsable_id) as responsable , " \
                              "ac.id, (SELECT ubi.nombre AS ubicacion FROM sagest_ubicacion ubi WHERE ubi.id = ac.ubicacion_id)," \
                              "(SELECT est.nombre AS estado FROM sagest_estadoproducto est WHERE est.id = ac.estado_id) " \
                              "from sagest_activofijo ac,sagest_catalogobien cat,sagest_gruposcategoria gru " \
                              "where ac.catalogo_id=cat.id  and cat.equipoelectronico=True and cat.grupo_id=gru.id " \
                              "and ac.statusactivo=1 and ac.status=True and (ac.archivobaja='' or ac.archivobaja isnull) " \
                              "AND ac.fechaingreso >= '{}' AND ac.fechaingreso <= '{}' AND NOT ac.id IN {} ORDER BY ac.fechaingreso ASC".format(
                            fdesde, fhasta, str(tuple(values_ac)))
                    else:
                        sql = "select ac.codigointerno,ac.codigogobierno,ac.fechaingreso,age (current_date, ac.fechaingreso) || '' tiempo, " \
                              "cast(extract(year from age (current_date, ac.fechaingreso))*12 + extract(month from age (ac.fechaingreso)) as int)/12 as numanios " \
                              ",cat.descripcion as catalogo,ac.descripcion,ac.modelo,ac.marca," \
                              "(select perso.apellido1 || ' ' || perso.apellido2 || ' ' || perso.nombres from sga_persona perso where id=ac.responsable_id) as responsable , " \
                              "ac.id, (SELECT ubi.nombre AS ubicacion FROM sagest_ubicacion ubi WHERE ubi.id = ac.ubicacion_id)," \
                              "(SELECT est.nombre AS estado FROM sagest_estadoproducto est WHERE est.id = ac.estado_id) " \
                              "from sagest_activofijo ac,sagest_catalogobien cat,sagest_gruposcategoria gru " \
                              "where ac.catalogo_id=cat.id  and cat.equipoelectronico=True and cat.grupo_id=gru.id " \
                              "and ac.statusactivo=1 and ac.status=True and gru.id=" + lista1 + " and (ac.archivobaja='' or ac.archivobaja isnull) " \
                                                                                                "AND ac.fechaingreso >= '{}' AND ac.fechaingreso <= '{}' AND NOT ac.id IN {} ORDER BY ac.fechaingreso ASC".format(
                            fdesde, fhasta, str(tuple(values_ac)))
                    cursor = connection.cursor()
                    cursor.execute(sql)
                    results = cursor.fetchall()
                    data['results'] = results
                    data['hoy'] = str(datetime.now().date())
                    # return conviert_html_to_pdf_name(
                    #     'inventario_activofijo/informes/activotecnologicopdf.html',
                    #     {'pagesize': 'A4', 'data': data,}, nombrearchivo)
                    return conviert_html_to_pdf(
                        'inventario_activofijo/informes/activotecnologicopdf.html',
                        {
                            'pagesize': 'A4',
                            'data': data,
                        },
                    )
                except Exception as ex:
                    pass

            elif action == 'pdflistadoactivosinactios':
                try:
                    data['title'] = nombrearchivo = u'Informe de Activos Tecnológicos'
                    fdesde, fhasta = request.GET.get('fechadesde', ''), request.GET.get('fechahasta', '')
                    data['desde'] = str(fdesde)
                    data['hasta'] = str(fhasta)
                    lista1 = '0'
                    if 'cadenatexto' in request.GET:
                        lista1 = request.GET['cadenatexto']
                    values_ac = InformeActivoBaja.objects.values('activofijo__id').filter(status=True,
                                                                                          activofijo__catalogo__equipoelectronico=True,
                                                                                          fecha_creacion__gte=fdesde,
                                                                                          fecha_creacion__lte=fhasta).values_list(
                        'activofijo__id', flat=True)
                    if not values_ac:
                        values_ac = [0]
                    if lista1 == '0':
                        sql = "select ac.codigointerno,ac.codigogobierno,ac.fechaingreso,age (current_date, ac.fechaingreso) || '' tiempo, " \
                              "cast(extract(year from age (current_date, ac.fechaingreso))*12 + extract(month from age (ac.fechaingreso)) as int)/12 as numanios " \
                              ",cat.descripcion as catalogo,ac.descripcion,ac.modelo,ac.marca," \
                              "(select perso.apellido1 || ' ' || perso.apellido2 || ' ' || perso.nombres from sga_persona perso where id=ac.responsable_id) as responsable , " \
                              "ac.id, (SELECT ubi.nombre AS ubicacion FROM sagest_ubicacion ubi WHERE ubi.id = ac.ubicacion_id)," \
                              "(SELECT est.nombre AS estado FROM sagest_estadoproducto est WHERE est.id = ac.estado_id), " \
                              "(SELECT DATE(fecha_creacion) FROM sagest_informeactivobaja WHERE activofijo_id=ac.id) AS fecha_salida " \
                              "from sagest_activofijo ac,sagest_catalogobien cat,sagest_gruposcategoria gru " \
                              "where ac.catalogo_id=cat.id  and cat.equipoelectronico=True and cat.grupo_id=gru.id " \
                              "and ac.statusactivo=1 and ac.status=True and (ac.archivobaja='' or ac.archivobaja isnull) " \
                              "AND ac.id IN {} ORDER BY fecha_salida ASC".format(str(tuple(values_ac)))
                    else:
                        sql = "select ac.codigointerno,ac.codigogobierno,ac.fechaingreso,age (current_date, ac.fechaingreso) || '' tiempo, " \
                              "cast(extract(year from age (current_date, ac.fechaingreso))*12 + extract(month from age (ac.fechaingreso)) as int)/12 as numanios " \
                              ",cat.descripcion as catalogo,ac.descripcion,ac.modelo,ac.marca," \
                              "(select perso.apellido1 || ' ' || perso.apellido2 || ' ' || perso.nombres from sga_persona perso where id=ac.responsable_id) as responsable , " \
                              "ac.id, (SELECT ubi.nombre AS ubicacion FROM sagest_ubicacion ubi WHERE ubi.id = ac.ubicacion_id)," \
                              "(SELECT est.nombre AS estado FROM sagest_estadoproducto est WHERE est.id = ac.estado_id), " \
                              "(SELECT DATE(fecha_creacion) FROM sagest_informeactivobaja WHERE activofijo_id=ac.id) AS fecha_salida " \
                              "from sagest_activofijo ac,sagest_catalogobien cat,sagest_gruposcategoria gru " \
                              "where ac.catalogo_id=cat.id  and cat.equipoelectronico=True and cat.grupo_id=gru.id " \
                              "and ac.statusactivo=1 and ac.status=True and gru.id=" + lista1 + " and (ac.archivobaja='' or ac.archivobaja isnull) " \
                                                                                                "AND  ac.id IN {} ORDER BY fecha_salida ASC".format(
                            str(tuple(values_ac)))
                    # cursor = connection.cursor()
                    # cursor.execute(sql)
                    # results = cursor.fetchall()
                    results = ActivoFijo.objects.filter(status=True, statusactivo=2, catalogo__equipoelectronico=True, catalogo__clasificado=True, catalogo__status=True, id__in=values_ac)
                    data['results'] = results
                    data['hoy'] = str(datetime.now().date())
                    # return conviert_html_to_pdf_name(
                    #     'inventario_activofijo/informes/activotecnologicopdf.html',
                    #     {'pagesize': 'A4', 'data': data,}, nombrearchivo)
                    return conviert_html_to_pdf(
                        'inventario_activofijo/informes/bajasactivotecnologicopdf.html',
                        {
                            'pagesize': 'A4',
                            'data': data,
                        },
                    )
                except Exception as ex:
                    pass

            elif action == 'pdflistadoactivosinactivos':
                try:
                    data['title'] = nombrearchivo = u'Informe de Activos Tecnológicos'
                    results = ActivoFijo.objects.filter(status=True, statusactivo=2, catalogo__equipoelectronico=True, catalogo__clasificado=True, catalogo__status=True)
                    data['results'] = results
                    data['hoy'] = str(datetime.now().date())
                    # return conviert_html_to_pdf_name(
                    #     'inventario_activofijo/informes/activotecnologicopdf.html',
                    #     {'pagesize': 'A4', 'data': data,}, nombrearchivo)
                    return conviert_html_to_pdf(
                        'inventario_activofijo/informes/bajasactivotecnologicopdf.html',
                        {
                            'pagesize': 'A4',
                            'data': data,
                        },
                    )
                except Exception as ex:
                    pass

            elif action == 'pdflistadoactivostodos':
                try:
                    data['title'] = nombrearchivo = u'Informe de Activos Tecnológicos'
                    lista1 = '0'
                    if 'cadenatexto' in request.GET:
                        lista1 = request.GET['cadenatexto']
                    values_ac = InformeActivoBaja.objects.values('activofijo__id').filter(status=True).values_list(
                        'activofijo__id', flat=True)
                    if lista1 == '0':
                        sql = "select ac.codigointerno,ac.codigogobierno,ac.fechaingreso,age (current_date, ac.fechaingreso) || '' tiempo, " \
                              "cast(extract(year from age (current_date, ac.fechaingreso))*12 + extract(month from age (ac.fechaingreso)) as int)/12 as numanios " \
                              ",cat.descripcion as catalogo,ac.descripcion,ac.modelo,ac.marca," \
                              "(select perso.apellido1 || ' ' || perso.apellido2 || ' ' || perso.nombres from sga_persona perso where id=ac.responsable_id) as responsable , " \
                              "ac.id, (SELECT ubi.nombre AS ubicacion FROM sagest_ubicacion ubi WHERE ubi.id = ac.ubicacion_id)," \
                              "(SELECT est.nombre AS estado FROM sagest_estadoproducto est WHERE est.id = ac.estado_id) " \
                              "from sagest_activofijo ac,sagest_catalogobien cat,sagest_gruposcategoria gru " \
                              "where ac.catalogo_id=cat.id  and cat.equipoelectronico=True and cat.grupo_id=gru.id " \
                              "and ac.statusactivo=1 and cat.clasificado=True and ac.status=True and (ac.archivobaja='' or ac.archivobaja isnull) " \
                              "AND NOT ac.id IN {} ORDER BY ac.fechaingreso ASC".format(str(tuple(values_ac)))
                    else:
                        sql = "select ac.codigointerno,ac.codigogobierno,ac.fechaingreso,age (current_date, ac.fechaingreso) || '' tiempo, " \
                              "cast(extract(year from age (current_date, ac.fechaingreso))*12 + extract(month from age (ac.fechaingreso)) as int)/12 as numanios " \
                              ",cat.descripcion as catalogo,ac.descripcion,ac.modelo,ac.marca," \
                              "(select perso.apellido1 || ' ' || perso.apellido2 || ' ' || perso.nombres from sga_persona perso where id=ac.responsable_id) as responsable , " \
                              "ac.id, (SELECT ubi.nombre AS ubicacion FROM sagest_ubicacion ubi WHERE ubi.id = ac.ubicacion_id)," \
                              "(SELECT est.nombre AS estado FROM sagest_estadoproducto est WHERE est.id = ac.estado_id) " \
                              "from sagest_activofijo ac,sagest_catalogobien cat,sagest_gruposcategoria gru " \
                              "where ac.catalogo_id=cat.id  and cat.equipoelectronico=True and cat.grupo_id=gru.id " \
                              "and ac.statusactivo=1 and ac.status=True and gru.id=" + lista1 + " and (ac.archivobaja='' or ac.archivobaja isnull) " \
                                                                                                "AND NOT ac.id IN {} ORDER BY ac.fechaingreso ASC".format(
                            str(tuple(values_ac)))
                    cursor = connection.cursor()
                    cursor.execute(sql)
                    results = cursor.fetchall()
                    data['results'] = results
                    data['hoy'] = str(datetime.now().date())
                    # return conviert_html_to_pdf_name(
                    #     'inventario_activofijo/informes/activotecnologicopdf.html',
                    #     {'pagesize': 'A4', 'data': data,}, nombrearchivo)
                    return conviert_html_to_pdf(
                        'inventario_activofijo/informes/activotecnologicopdf.html',
                        {
                            'pagesize': 'A4',
                            'data': data,
                        },
                    )
                except Exception as ex:
                    pass

            elif action == 'pdflistadoactivosinactiostodos':
                try:
                    data['title'] = nombrearchivo = u'Informe de Activos Tecnológicos'
                    fdesde, fhasta = request.GET.get('fechadesde', ''), request.GET.get('fechahasta', '')
                    data['desde'] = str(fdesde)
                    data['hasta'] = str(fhasta)
                    lista1 = '0'
                    if 'cadenatexto' in request.GET:
                        lista1 = request.GET['cadenatexto']
                    values_ac = InformeActivoBaja.objects.values('activofijo__id').filter(status=True,
                                                                                          activofijo__catalogo__equipoelectronico=True,
                                                                                          fecha_creacion__gte=fdesde,
                                                                                          fecha_creacion__lte=fhasta).values_list(
                        'activofijo__id', flat=True)
                    if not values_ac:
                        values_ac = [0]
                    if lista1 == '0':
                        sql = "select ac.codigointerno,ac.codigogobierno,ac.fechaingreso,age (current_date, ac.fechaingreso) || '' tiempo, " \
                              "cast(extract(year from age (current_date, ac.fechaingreso))*12 + extract(month from age (ac.fechaingreso)) as int)/12 as numanios " \
                              ",cat.descripcion as catalogo,ac.descripcion,ac.modelo,ac.marca," \
                              "(select perso.apellido1 || ' ' || perso.apellido2 || ' ' || perso.nombres from sga_persona perso where id=ac.responsable_id) as responsable , " \
                              "ac.id, (SELECT ubi.nombre AS ubicacion FROM sagest_ubicacion ubi WHERE ubi.id = ac.ubicacion_id)," \
                              "(SELECT est.nombre AS estado FROM sagest_estadoproducto est WHERE est.id = ac.estado_id), " \
                              "(SELECT DATE(fecha_creacion) FROM sagest_informeactivobaja WHERE activofijo_id=ac.id) AS fecha_salida " \
                              "from sagest_activofijo ac,sagest_catalogobien cat,sagest_gruposcategoria gru " \
                              "where ac.catalogo_id=cat.id  and cat.equipoelectronico=True and cat.grupo_id=gru.id " \
                              "and ac.statusactivo=1 and ac.status=True and (ac.archivobaja='' or ac.archivobaja isnull) " \
                              "AND ac.id IN {} ORDER BY fecha_salida ASC".format(str(tuple(values_ac)))
                    else:
                        sql = "select ac.codigointerno,ac.codigogobierno,ac.fechaingreso,age (current_date, ac.fechaingreso) || '' tiempo, " \
                              "cast(extract(year from age (current_date, ac.fechaingreso))*12 + extract(month from age (ac.fechaingreso)) as int)/12 as numanios " \
                              ",cat.descripcion as catalogo,ac.descripcion,ac.modelo,ac.marca," \
                              "(select perso.apellido1 || ' ' || perso.apellido2 || ' ' || perso.nombres from sga_persona perso where id=ac.responsable_id) as responsable , " \
                              "ac.id, (SELECT ubi.nombre AS ubicacion FROM sagest_ubicacion ubi WHERE ubi.id = ac.ubicacion_id)," \
                              "(SELECT est.nombre AS estado FROM sagest_estadoproducto est WHERE est.id = ac.estado_id), " \
                              "(SELECT DATE(fecha_creacion) FROM sagest_informeactivobaja WHERE activofijo_id=ac.id) AS fecha_salida " \
                              "from sagest_activofijo ac,sagest_catalogobien cat,sagest_gruposcategoria gru " \
                              "where ac.catalogo_id=cat.id  and cat.equipoelectronico=True and cat.grupo_id=gru.id " \
                              "and ac.statusactivo=1 and ac.status=True and gru.id=" + lista1 + " and (ac.archivobaja='' or ac.archivobaja isnull) " \
                                                                                                "AND  ac.id IN {} ORDER BY fecha_salida ASC".format(
                            str(tuple(values_ac)))
                    cursor = connection.cursor()
                    cursor.execute(sql)
                    results = cursor.fetchall()
                    data['results'] = results
                    data['hoy'] = str(datetime.now().date())
                    # return conviert_html_to_pdf_name(
                    #     'inventario_activofijo/informes/activotecnologicopdf.html',
                    #     {'pagesize': 'A4', 'data': data,}, nombrearchivo)
                    return conviert_html_to_pdf(
                        'inventario_activofijo/informes/bajasactivotecnologicopdf.html',
                        {
                            'pagesize': 'A4',
                            'data': data,
                        },
                    )
                except Exception as ex:
                    pass

            elif action == 'editacta':
                try:
                    data['title'] = u'Editar Acta'
                    data['acta'] = acta = TraspasoActivo.objects.get(pk=request.GET['id'])
                    detalle = acta.detalletraspasoactivo_set.all()[0]
                    actas = acta.detalletraspasoactivo_set.all()
                    form = ActasForm(initial={
                        'numero': acta.numero,
                        'fecha': acta.fecha_creacion,
                        'tipobien': detalle.activo.catalogo.tipobien,
                        'tipocomprobante': detalle.activo.tipocomprobante,
                        'origeningreso': detalle.activo.origeningreso,
                        'proveedor': acta.proveedor,
                        'responsable': acta.usuariobienrecibe,
                        'custodio': acta.custodiobienrecibe,
                        'ubicacion': acta.ubicacionbienrecibe,
                        'observacion': acta.observacion,
                    })
                    form.editar()
                    data['form'] = form
                    paging = MiPaginador(actas, 10)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['detalles'] = page.object_list
                    data['usuario'] = request.user
                    return render(request, 'inventario_activofijo/editacta.html', data)
                except Exception as ex:
                    pass

            elif action == 'consultacatalogo':
                try:
                    data['title'] = u'Consultas de activos por catálogos'
                    data['form'] = ConsultacatalogoForm()
                    data['permite_modificar'] = False
                    data['reporte_0'] = obtener_reporte('consulta_activos_catalogo')
                    return render(request, 'inventario_activofijo/consultacatalogo.html', data)
                except Exception as ex:
                    pass

            elif action == 'consultausuario':
                try:
                    data['title'] = u'Consultas de activos por usuario'
                    data['form'] = ConsultausuarioForm()
                    data['permite_modificar'] = False
                    data['reporte_0'] = obtener_reporte('consulta_activos_usuario')
                    return render(request, 'inventario_activofijo/consultausuario.html', data)
                except Exception as ex:
                    pass

            elif action == 'asignacion':
                try:
                    data['title'] = u'Asignar usuario y custodio'
                    data['traspasos'] = TraspasoActivo.objects.filter(tipo=1)
                    return render(request, "inventario_activofijo/asignacion.html", data)
                except Exception as ex:
                    pass

            elif action == 'asignar':
                try:
                    data['title'] = u'Asignar activos'
                    data['form'] = AsignacionActivoForm()
                    data['activos'] = ActivoFijo.objects.filter(detalletraspasoactivo__isnull=True)
                    return render(request, "inventario_activofijo/asignar.html", data)
                except Exception as ex:
                    pass

            elif action == 'historial':
                try:
                    data['title'] = u'Historial del Bien'
                    data['activo'] = activo = ActivoFijo.objects.get(pk=int(request.GET['id']))
                    # data["detalles"] = activo.detalletraspasoactivo_set.filter(seleccionado=True).order_by('-codigotraspaso__fecha')
                    data["detalles"] = activo.detalletraspasoactivo_set.filter(status=True).order_by(
                        '-codigotraspaso__fecha')
                    data['reporte_0'] = obtener_reporte('historial_activos')
                    data['usuario'] = request.user
                    return render(request, "inventario_activofijo/historial.html", data)
                except Exception as ex:
                    pass

            elif action == 'constataciones':
                try:
                    data['title'] = u'Constataciones del Bien'
                    data['activo'] = activo = ActivoFijo.objects.get(pk=int(request.GET['id']))
                    data["detalles_pert"] = activo.detalleconstatacionfisica_set.filter(perteneceusuario=True).order_by(
                        '-codigoconstatacion__numero')
                    data['detallenopert'] = activo.detalleconstatacionfisica_set.filter(perteneceusuario=False)
                    data['detallenotras'] = activo.detalleconstatacionfisica_set.filter(requieretraspaso=True)
                    data['reporte_0'] = obtener_reporte('historial_activos')
                    data['usuario'] = request.user
                    return render(request, "inventario_activofijo/detalle_constataciones.html", data)
                except Exception as ex:
                    pass

            elif action == 'importar':
                try:
                    data['title'] = u'Importar lista de Activos'
                    data['form'] = ImportarArchivoXLSForm()
                    return render(request, "inventario_activofijo/importar.html", data)
                except Exception as ex:
                    pass

            elif action == 'exportar':
                try:
                    data['title'] = u'Exportar lista de Activos'
                    data['form'] = ExportacionForm()
                    return render(request, "inventario_activofijo/exportar.html", data)
                except Exception as ex:
                    pass

            elif action == 'subir':
                try:
                    data['title'] = u'Subir Archivo'
                    data['importacion'] = importacion = ArchivoActivoFijo.objects.get(pk=request.GET['id'])
                    form = ImportarArchivoXLSForm()
                    form.editar()
                    data['form'] = form
                    return render(request, "inventario_activofijo/subir.html", data)
                except Exception as ex:
                    pass

            elif action == 'importaciones':
                try:
                    data['title'] = u'Importaciones'
                    search = None
                    ids = None
                    if 's' in request.GET:
                        search = request.GET['s']
                        importacion = ArchivoActivoFijo.objects.filter(Q(tipobien__nombre__icontains=search) |
                                                                       Q(nombre__icontains=search)).order_by('-id')
                    elif 'id' in request.GET:
                        ids = request.GET['id']
                        importacion = ArchivoActivoFijo.objects.filter(id=ids).order_by('-id')
                    else:
                        importacion = ArchivoActivoFijo.objects.all().order_by('-id')
                    paging = MiPaginador(importacion, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['reporte_0'] = obtener_reporte('acta_constatacion')
                    data['importaciones'] = page.object_list
                    return render(request, "inventario_activofijo/importaciones.html", data)
                except Exception as ex:
                    pass

            elif action == 'exportaciones':
                try:
                    data['title'] = u'Exportaciones'
                    search = None
                    ids = None
                    if 's' in request.GET:
                        search = request.GET['s']
                        exportaciones = ExportacionesActivos.objects.filter(
                            Q(cuentacontable__descripcion__icontains=search))
                    elif 'id' in request.GET:
                        ids = request.GET['id']
                        exportaciones = ExportacionesActivos.objects.filter(id=ids)
                    else:
                        exportaciones = ExportacionesActivos.objects.all()
                    paging = MiPaginador(exportaciones, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['reporte_0'] = obtener_reporte('acta_constatacion')
                    data['exportaciones'] = page.object_list
                    return render(request, "inventario_activofijo/exportaciones.html", data)
                except Exception as ex:
                    pass

            elif action == 'addtraspaso':
                try:
                    puede_realizar_accion(request, 'sagest.puede_modificar_traspasos')
                    data['title'] = u'Nuevo Traspaso'
                    form = TraspasoActivoForm()
                    form.adicionar()
                    data['form'] = form
                    return render(request, "inventario_activofijo/addtraspaso.html", data)
                except Exception as ex:
                    pass

            elif action == 'addtraspasocustodio':
                try:
                    puede_realizar_accion(request, 'sagest.puede_modificar_traspasos')
                    data['title'] = u'Nuevo Traspaso de Custodios'
                    form = TraspasoActivoCustodioForm()
                    form.adicionarcustodio()
                    data['form'] = form
                    return render(request, "inventario_activofijo/addtraspasocustodio.html", data)
                except Exception as ex:
                    pass

            elif action == 'addbaja':
                try:
                    puede_realizar_accion(request, 'sagest.puede_ingresar_compras')
                    data['title'] = u'Nueva Baja de Activos'
                    data['form'] = BajaActivoForm()
                    data['form2'] = DetalleBajaActivoForm()
                    return render(request, "inventario_activofijo/addbaja.html", data)
                except Exception as ex:
                    pass

            elif action == 'addtipobaja':
                try:
                    data['title'] = u'Nuevo Tipo Baja'
                    data['form'] = TipoBajaForm()
                    return render(request, "inventario_activofijo/addtipobaja.html", data)
                except Exception as ex:
                    pass

            elif action == 'addtraslado':
                try:
                    data['title'] = u'Nuevo Traslado de Activos'
                    data['form'] = TrasladoMantenimientoForm()
                    return render(request, "inventario_activofijo/addtraslado.html", data)
                except Exception as ex:
                    pass

            elif action == 'edittraslado':
                try:
                    data['title'] = u'Editar Traslado de Activos'
                    data['traslado'] = traslado = TrasladoMantenimiento.objects.get(pk=request.GET['id'])
                    form = TrasladoMantenimientoForm(initial={'departamentosolicita': traslado.departamentosolicita,
                                                              'asistentelogistica': traslado.asistentelogistica,
                                                              'taller': traslado.taller,
                                                              'observacion': traslado.observacion,
                                                              'administradorcontrato': traslado.administradorcontrato,
                                                              'usuariobienes': traslado.usuariobienes})
                    form.editar()
                    data['form'] = form
                    return render(request, "inventario_activofijo/edittraslado.html", data)
                except Exception as ex:
                    pass

            elif action == 'addconstatacion':
                try:
                    data['title'] = u'Nueva Constatación'
                    form = ConstatacionForm()
                    form.adicionar()
                    data['form'] = form
                    return render(request, "inventario_activofijo/addconstatacion.html", data)
                except Exception as ex:
                    pass

            elif action == 'editcons':
                try:
                    data['title'] = u'Editar Constatación'
                    data['constatacion'] = constatacion = ConstatacionFisica.objects.get(pk=request.GET['id'])
                    form = ConstatacionForm(initial={'usuariobienes': constatacion.usuariobienes.id,
                                                     'ubicacionbienes': constatacion.ubicacionbienes,
                                                     'numero': constatacion.numero,
                                                     'observacion': constatacion.observacion})
                    form.editar(constatacion)
                    data['form'] = form
                    data['form3'] = DetalleNoIdentificadoForm()
                    return render(request, "inventario_activofijo/editconstatacion.html", data)
                except Exception as ex:
                    pass

            elif action == 'cambiousuario':
                try:
                    data['title'] = u'Editar Constatación'
                    data['constatacion'] = constatacion = ConstatacionFisica.objects.get(pk=request.GET['id'])
                    form = ConstatacionForm(initial={'usuariobienes': constatacion.usuariobienes.id,
                                                     'ubicacionbienes': constatacion.ubicacionbienes,
                                                     'numero': constatacion.numero,
                                                     'observacion': constatacion.observacion})
                    form.cambiar(constatacion)
                    data['form'] = form
                    return render(request, "inventario_activofijo/cambiousuario.html", data)
                except Exception as ex:
                    pass

            elif action == 'edittraspaso':
                try:
                    puede_realizar_accion(request, 'sagest.puede_modificar_traspasos')
                    data['title'] = u'Editar Traspaso de activos'
                    data['traspaso'] = traspaso = TraspasoActivo.objects.get(pk=request.GET['id'])
                    data['detalles'] = traspaso.detalletraspasoactivo_set.all().order_by('activo__codigointerno',
                                                                                         'activo__codigogobierno')
                    initial = model_to_dict(traspaso)
                    form = TraspasoActivoForm(initial=initial)
                    form.editar()
                    data['form'] = form
                    return render(request, "inventario_activofijo/edittraspaso.html", data)
                except Exception as ex:
                    pass

            elif action == 'edittraspasocustodio':
                try:
                    puede_realizar_accion(request, 'sagest.puede_modificar_traspasos')
                    data['title'] = u'Editar Traspaso de activos'
                    data['traspaso'] = traspaso = TraspasoActivo.objects.get(pk=request.GET['id'])
                    data['detalles'] = traspaso.detalletraspasoactivo_set.all().order_by('activo__codigointerno',
                                                                                         'activo__codigogobierno')
                    initial = model_to_dict(traspaso)
                    form = TraspasoActivoCustodioForm(initial=initial)
                    form.editarcustodio()
                    data['form'] = form
                    return render(request, "inventario_activofijo/edittraspasocustodio.html", data)
                except Exception as ex:
                    pass

            elif action == 'editbaja':
                try:
                    puede_realizar_accion(request, 'sagest.puede_modificar_bajas')
                    data['title'] = u'Editar Baja de activos'
                    data['baja'] = baja = BajaActivo.objects.get(pk=request.GET['id'])
                    data['detalles'] = baja.detallebajaactivo_set.all().order_by('activo__codigointerno',
                                                                                 'activo__codigogobierno')
                    initial = model_to_dict(baja)
                    form = BajaActivoForm(initial=initial)
                    form.adicionar()
                    data['form'] = form
                    return render(request, "inventario_activofijo/editbaja.html", data)
                except Exception as ex:
                    pass

            elif action == 'edittipobaja':
                try:
                    data['title'] = u'Editar Tipo Baja'
                    data['tipobaja'] = tipobaja = TipoBaja.objects.get(pk=request.GET['id'])
                    initial = model_to_dict(tipobaja)
                    form = TipoBajaForm(initial=initial)
                    data['form'] = form
                    return render(request, "inventario_activofijo/edittipobaja.html", data)
                except Exception as ex:
                    pass

            elif action == 'histconstatacion':
                try:
                    data['title'] = u'Constatación Física'
                    search = None
                    ids = None
                    constatacion = ConstatacionFisica.objects.filter(status=True).order_by('estado', '-numero')
                    if 's' in request.GET:
                        search = request.GET['s']
                        ss = search.split(' ')
                        if len(ss) == 1:
                            constatacion = constatacion.filter(Q(numero__icontains=search) |
                                                               Q(usuariobienes__nombres__icontains=search) |
                                                               Q(usuariobienes__apellido1__icontains=search) |
                                                               Q(usuariobienes__apellido2__icontains=search) |
                                                               Q(ubicacionbienes__nombre__icontains=search)).distinct()
                        else:
                            constatacion = constatacion.filter(Q(usuariobienes__apellido1__icontains=ss[0]) &
                                                               Q(usuariobienes__apellido2__icontains=ss[1])).distinct()

                    elif 'id' in request.GET:
                        ids = int(request.GET['id'])
                        constatacion = constatacion.filter(id=ids)
                    paging = MiPaginador(constatacion, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['reporte_0'] = obtener_reporte('acta_constatacion')
                    data['constataciones'] = page.object_list
                    data['usuario'] = request.user
                    return render(request, "inventario_activofijo/constataciones.html", data)
                except Exception as ex:
                    pass

            elif action == 'histconstatacionencontrados':
                try:
                    data['title'] = u'Constatación Física Bienes Encontrados'
                    search = None
                    ids = None
                    # constatacion = ConstatacionFisica.objects.filter(status=True).order_by('estado', '-numero')
                    if 's' in request.GET:
                        search = request.GET['s']
                        ss = search.split(' ')
                        if len(ss) == 1:
                            constatacion = DetalleConstatacionFisica.objects.filter(
                                Q(activo__codigogobierno__icontains=search) |
                                Q(activo__codigointerno__icontains=search), encontrado=True, enuso=True, status=True,
                                codigoconstatacion__status=True).distinct().order_by('codigoconstatacion__estado',
                                                                                     '-codigoconstatacion__numero')
                        else:
                            constatacion = DetalleConstatacionFisica.objects.filter(encontrado=True, enuso=True,
                                                                                    status=True,
                                                                                    codigoconstatacion__status=True).distinct().order_by(
                                'codigoconstatacion__estado', '-codigoconstatacion__numero')

                    # elif 'id' in request.GET:
                    #     ids = int(request.GET['id'])
                    #     constatacion = constatacion.filter(id=ids)
                    else:
                        constatacion = DetalleConstatacionFisica.objects.filter(encontrado=True, enuso=True,
                                                                                status=True,
                                                                                codigoconstatacion__status=True).distinct().order_by(
                            'codigoconstatacion__estado', '-codigoconstatacion__numero')
                    paging = MiPaginador(constatacion, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['reporte_0'] = obtener_reporte('acta_constatacion')
                    data['constataciones'] = page.object_list
                    data['usuario'] = request.user
                    return render(request, "inventario_activofijo/constatacionesencontrados.html", data)
                except Exception as ex:
                    pass

            elif action == 'finalizarcons':
                try:
                    data['title'] = u'Confirmar finalizar constatación'
                    data['constatacion'] = ConstatacionFisica.objects.get(pk=int(request.GET['id']))
                    return render(request, "inventario_activofijo/finalizar.html", data)
                except:
                    pass

            elif action == 'confirmarexpo':
                try:
                    data['title'] = u'Confirmar Activos Como subidos a Gobierno'
                    data['exportacion'] = ExportacionesActivos.objects.get(pk=int(request.GET['id']))
                    return render(request, "inventario_activofijo/confirmaexpo.html", data)
                except:
                    pass

            elif action == 'finalizartraslado':
                try:
                    data['title'] = u'Confirmar finalizar traslado'
                    data['traslado'] = TrasladoMantenimiento.objects.get(pk=int(request.GET['id']))
                    return render(request, "inventario_activofijo/finalizartraslado.html", data)
                except:
                    pass

            elif action == 'finalizartraspaso':
                try:
                    data['title'] = u'Confirmar finalizar traspaso'
                    data['traspaso'] = TraspasoActivo.objects.get(pk=int(request.GET['id']))
                    return render(request, "inventario_activofijo/finalizartraspaso.html", data)
                except:
                    pass

            elif action == 'abrirtraspaso':
                try:
                    data['title'] = u'Confirmar reabrir traspaso'
                    data['traspaso'] = TraspasoActivo.objects.get(pk=int(request.GET['id']))
                    return render(request, "inventario_activofijo/abrirtraspaso.html", data)
                except:
                    pass

            elif action == 'eliminartraspaso':
                try:
                    data['title'] = u'Confirmar eliminar traspaso'
                    data['traspaso'] = TraspasoActivo.objects.get(pk=int(request.GET['id']))
                    return render(request, "inventario_activofijo/eliminartraspaso.html", data)
                except:
                    pass

            elif action == 'deleteexpo':
                try:
                    data['title'] = u'Confirmar eliminar exportacion'
                    data['exportacion'] = ExportacionesActivos.objects.get(pk=int(request.GET['id']))
                    return render(request, "inventario_activofijo/deleteexpo.html", data)
                except:
                    pass

            elif action == 'deleteimpo':
                try:
                    data['title'] = u'Confirmar eliminar importacion'
                    data['importacion'] = ArchivoActivoFijo.objects.get(pk=int(request.GET['id']))
                    return render(request, "inventario_activofijo/deleteimpo.html", data)
                except:
                    pass

            elif action == 'eliminartraslado':
                try:
                    data['title'] = u'Confirmar eliminar traslado mantenimiento'
                    data['traslado'] = TrasladoMantenimiento.objects.get(pk=int(request.GET['id']))
                    return render(request, "inventario_activofijo/eliminartraslado.html", data)
                except:
                    pass

            elif action == 'eliminarconstatacion':
                try:
                    data['title'] = u'Confirmar eliminar constatacion'
                    data['constatacion'] = ConstatacionFisica.objects.get(pk=int(request.GET['id']))
                    return render(request, "inventario_activofijo/eliminarconstatacion.html", data)
                except:
                    pass

            elif action == 'eliminarbaja':
                try:
                    data['title'] = u'Confirmar eliminar baja'
                    data['baja'] = BajaActivo.objects.get(pk=int(request.GET['id']))
                    return render(request, "inventario_activofijo/eliminarbaja.html", data)
                except:
                    pass

            elif action == 'eliminartipobaja':
                try:
                    data['title'] = u'Confirmar eliminar tipo baja'
                    data['tipobaja'] = TipoBaja.objects.get(pk=int(request.GET['id']))
                    return render(request, "inventario_activofijo/eliminartipobaja.html", data)
                except:
                    pass

            elif action == 'finalizarbaja':
                try:
                    data['title'] = u'Confirmar finalizar baja de activos'
                    data['baja'] = BajaActivo.objects.get(pk=int(request.GET['id']))
                    return render(request, "inventario_activofijo/finalizarbaja.html", data)
                except:
                    pass

            elif action == 'finalizarmantenimiento':
                try:
                    data['title'] = u'Confirmar finalizar de la mantenimiento'
                    data['mantenimiento'] = DetalleTrasladoMantenimiento.objects.get(pk=int(request.GET['id']))
                    return render(request, "inventario_activofijo/finalizarman.html", data)
                except:
                    pass

            elif action == 'finalizaracta':
                try:
                    data['title'] = u'Confirmar finalización de esta acta de Entrega'
                    data['acta'] = TraspasoActivo.objects.get(pk=int(request.GET['id']))
                    return render(request, "inventario_activofijo/finalizaracta.html", data)
                except:
                    pass

            elif action == 'histbajas':
                try:
                    data['title'] = u'Bajas de Activos'
                    search = None
                    ids = None
                    baja = BajaActivo.objects.filter(status=True).order_by('estado', '-numero')
                    if 's' in request.GET:
                        search = request.GET['s']
                        ss = search.split(' ')
                        if len(ss) == 1:
                            baja = baja.filter(Q(numero__icontains=search) |
                                               Q(solicitante__nombres__icontains=search) |
                                               Q(solicitante__apellido1__icontains=search) |
                                               Q(solicitante__apellido2__icontains=search) |
                                               Q(usuariobienentrega__nombres__icontains=search) |
                                               Q(usuariobienentrega__apellido1__icontains=search) |
                                               Q(usuariobienentrega__apellido2__icontains=search) |
                                               Q(ubicacionbienentrega__nombre__icontains=search)).distinct().order_by(
                                'estado', '-numero')
                        else:
                            baja = baja.filter(Q(usuariobienentrega__apellido1__icontains=ss[0]) &
                                               Q(usuariobienentrega__apellido2__icontains=ss[1])).distinct().order_by(
                                'estado', '-numero')
                    elif 'id' in request.GET:
                        ids = request.GET['id']
                        baja = baja.filter(id=ids).order_by('estado', '-numero')
                    paging = MiPaginador(baja, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['bajas'] = page.object_list
                    data['reporte_0'] = obtener_reporte('bajas_activo')
                    data['usuario'] = request.user
                    return render(request, "inventario_activofijo/bajas.html", data)
                except Exception as ex:
                    pass

            elif action == 'tiposbajas':
                try:
                    data['title'] = u'Mantenimiento Tipo Bajas'
                    search = None
                    ids = None
                    tipobaja = TipoBaja.objects.filter(status=True).order_by('id')
                    if 's' in request.GET:
                        search = request.GET['s']
                        ss = search.split(' ')
                        if len(ss) == 1:
                            tipobaja = tipobaja.filter(nombre__icontains=search).distinct().order_by('id')
                        else:
                            tipobaja = tipobaja.filter(Q(nombre__icontains=ss[0]) &
                                                       Q(nombre__icontains=ss[1])).distinct().order_by('id')
                    elif 'id' in request.GET:
                        ids = request.GET['id']
                        tipobaja = tipobaja.filter(id=ids).order_by('id')
                    paging = MiPaginador(tipobaja, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['tipobajas'] = page.object_list
                    data['usuario'] = request.user
                    return render(request, "inventario_activofijo/tiposbajas.html", data)
                except Exception as ex:
                    pass

            elif action == 'addtarjeta':
                try:
                    puede_realizar_accion(request, 'sagest.puede_modificar_tarjeta')
                    data['title'] = u'Nuevo Detalle de Mantenimiento y Reparación.'
                    data['activo'] = activo = ActivoFijo.objects.get(pk=request.GET['id'])
                    data['tarjeta'] = tarjeta = activo.mi_tarjeta_control()
                    form = DetalleTarjetaControlForm(
                        initial={'codigobara': activo.codigogobierno if activo.codigogobierno else activo.codigointerno,
                                 'catalogo': activo.catalogo,
                                 'descripcion': activo.descripcion,
                                 'fechacompra': activo.fechaingreso})
                    form.agregar()
                    data['form'] = form
                    return render(request, "inventario_activofijo/addtarjeta.html", data)
                except Exception as ex:
                    pass

            elif action == 'edittarjeta':
                try:
                    puede_realizar_accion(request, 'sagest.puede_modificar_tarjeta')
                    data['title'] = u'Editar Envío a Mantenimiento.'
                    data['detalle'] = detalle = DetalleMantenimiento.objects.get(pk=request.GET['id'])
                    activo = detalle.tarjeta.activo
                    form = DetalleTarjetaControlForm(
                        initial={'codigobara': activo.codigogobierno if activo.codigogobierno else activo.codigointerno,
                                 'catalogo': activo.catalogo,
                                 'descripcion': activo.descripcion,
                                 'fechacompra': activo.fechaingreso,
                                 'fechaentrega': detalle.fechaentrega if detalle.fechaentrega else datetime.now().date(),
                                 'mantenimientorealizar': detalle.mantenimientorealizar,
                                 'observacion': detalle.observacion,
                                 'taller': detalle.taller})
                    form.agregar()
                    data['form'] = form
                    return render(request, "inventario_activofijo/edittarjeta.html", data)
                except Exception as ex:
                    pass

            elif action == 'ingresodemantenimiento':
                try:
                    puede_realizar_accion(request, 'sagest.puede_modificar_tarjeta')
                    data['title'] = u'Ingreso de Mantenimiento y Reparación.'
                    data['detalle'] = detalle = DetalleMantenimiento.objects.get(pk=request.GET['id'])
                    activo = detalle.tarjeta.activo
                    traslaso = None
                    if detalle.detalletrasladomantenimiento:
                        traslado = detalle.detalletrasladomantenimiento.codigotraslado
                    form = DetalleTarjetaControlForm(initial={'mantenimientorealizar': detalle.mantenimientorealizar,
                                                              'mantenimientorealizado': detalle.mantenimientorealizado,
                                                              'codigobara': activo.codigogobierno if activo.codigogobierno else activo.codigointerno,
                                                              'catalogo': activo.catalogo,
                                                              'descripcion': activo.descripcion,
                                                              'fechacompra': activo.fechaingreso,
                                                              'observacion': detalle.observacion,
                                                              'aplicagarantia': detalle.aplicagarantia,
                                                              'taller': detalle.taller,
                                                              'manodeobra': detalle.manodeobra,
                                                              'costomanodeobra': detalle.costomanodeobra,
                                                              'facturamanodeobra': detalle.facturamanodeobra,
                                                              'fechaentrega': detalle.fechaentrega})
                    form.ingreso()
                    data['form'] = form
                    return render(request, "inventario_activofijo/ingresomantenimiento.html", data)
                except Exception as ex:
                    pass

            elif action == 'histtraslados':
                try:
                    data['title'] = u'Traslados de activos a mantenimiento'
                    search = None
                    ids = None
                    traslado = TrasladoMantenimiento.objects.filter(status=True).order_by('estado', '-numero')
                    if 's' in request.GET:
                        search = request.GET['s']
                        ss = search.split(' ')
                        if len(ss) == 1:
                            traslado = traslado.filter(Q(numero__icontains=search) |
                                                       Q(detalletrasladomantenimiento__activo__codigointerno=search) |
                                                       Q(asistentelogistica__nombres__icontains=search) |
                                                       Q(asistentelogistica__apellido1__icontains=search) |
                                                       Q(asistentelogistica__apellido2__icontains=search) |
                                                       Q(usuariobienes__nombres__icontains=search) |
                                                       Q(usuariobienes__apellido1__icontains=search) |
                                                       Q(usuariobienes__apellido2__icontains=search) |
                                                       Q(departamentosolicita__nombre__icontains=search)).distinct().order_by(
                                'estado', '-numero')
                        else:
                            traslado = traslado.filter(Q(usuariobienes__apellido1__icontains=ss[0]) &
                                                       Q(usuariobienes__apellido2=ss[1]) |
                                                       Q(asistentelogistica__apellido1__icontains=ss[0]) &
                                                       Q(asistentelogistica__apellido2=ss[1])).distinct().order_by(
                                'estado', '-numero')
                    elif 'id' in request.GET:
                        ids = request.GET['id']
                        traslado = traslado.filter(id=ids).order_by('estado', '-numero')
                    paging = MiPaginador(traslado, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['traslados'] = page.object_list
                    data['reporte_0'] = obtener_reporte('traslado_activo')
                    data['usuario'] = request.user
                    return render(request, "inventario_activofijo/traslados.html", data)
                except Exception as ex:
                    pass

            elif action == 'movimientos':
                try:
                    data['title'] = u'Movimientos de Activos'
                    searchse = None
                    searchsr = None
                    search = None
                    ids = None
                    traspasos = TraspasoActivo.objects.filter(tipo=2, status=True).order_by('estado', '-numero')
                    if 'se' in request.GET or 'sr' in request.GET or 's' in request.GET:
                        if 'se' in request.GET:
                            searchse = request.GET['se'].strip()
                            ss = searchse.split(' ')
                            if len(ss) == 1:
                                traspasos = traspasos.filter(Q(usuariobienentrega__nombres__icontains=searchse) |
                                                             Q(usuariobienentrega__apellido1__icontains=searchse) |
                                                             Q(usuariobienentrega__apellido2__icontains=searchse),
                                                             detalletraspasoactivo__seleccionado=True).distinct().order_by(
                                    'estado', '-numero')
                            else:
                                traspasos = traspasos.filter(Q(usuariobienentrega__apellido1__icontains=ss[0]) & Q(
                                    usuariobienentrega__apellido2__icontains=ss[1]),
                                                             detalletraspasoactivo__seleccionado=True).distinct().order_by(
                                    'estado', '-numero')
                        if 's' in request.GET:
                            search = request.GET['s']
                            traspasos = traspasos.filter(Q(numero__icontains=search) |
                                                         Q(usuario_creacion__username=search) |
                                                         Q(detalletraspasoactivo__activo__codigointerno=search,
                                                           detalletraspasoactivo__seleccionado=True) |
                                                         Q(detalletraspasoactivo__activo__codigogobierno=search,
                                                           detalletraspasoactivo__seleccionado=True)).distinct().order_by(
                                'estado', '-numero')

                        if 'sr' in request.GET:
                            searchsr = request.GET['sr'].strip()
                            ss = searchsr.split(' ')
                            if len(ss) == 1:
                                traspasos = traspasos.filter(Q(usuariobienrecibe__nombres__icontains=searchsr) |
                                                             Q(usuariobienrecibe__apellido1__icontains=searchsr) |
                                                             Q(usuariobienrecibe__apellido2__icontains=searchsr),
                                                             detalletraspasoactivo__seleccionado=True).distinct().order_by(
                                    'estado', '-numero')
                            else:
                                traspasos = traspasos.filter(Q(usuariobienentrega__apellido1__icontains=ss[0]) & Q(
                                    usuariobienentrega__apellido2__icontains=ss[1]),
                                                             detalletraspasoactivo__seleccionado=True).distinct().order_by(
                                    'estado', '-numero')
                    elif 'id' in request.GET:
                        ids = int(request.GET['id'])
                        traspasos = traspasos.filter(id=ids).order_by('estado', '-numero')
                    paging = MiPaginador(traspasos, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['searchse'] = searchse if searchse else ""
                    data['searchsr'] = searchsr if searchsr else ""
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['traspasos'] = page.object_list
                    data['reporte_0'] = obtener_reporte('traspaso_activo')
                    data['reporte_1'] = obtener_reporte('traspaso_activo_solocustodio')
                    data['usuario'] = request.user
                    return render(request, "inventario_activofijo/movimientos.html", data)
                except Exception as ex:
                    pass

            elif action == 'actasentrega':
                try:
                    data['title'] = u'Actas de entrega y recepción de Activos'
                    searchse = None
                    searchsr = None
                    ids = None
                    if 'se' in request.GET or 'sr' in request.GET:
                        traspasos = TraspasoActivo.objects.filter(tipo=1)
                        if 'se' in request.GET:
                            searchse = request.GET['se'].strip()
                            ss = searchse.split(' ')
                            if len(ss) == 1:
                                traspasos = traspasos.filter(
                                    Q(usuariobienentrega__nombres__icontains=searchse, tipo=1) |
                                    Q(usuariobienentrega__apellido1__icontains=searchse, tipo=1) |
                                    Q(usuariobienentrega__apellido2__icontains=searchse, tipo=1)).order_by('-numero')
                            else:
                                traspasos = traspasos.filter(
                                    Q(usuariobienentrega__apellido1__icontains=ss[0], tipo=1) & Q(
                                        usuariobienentrega__apellido2__icontains=ss[1], tipo=1)).order_by('-numero')
                        if 'sr' in request.GET:
                            searchsr = request.GET['sr'].strip()
                            ss = searchsr.split(' ')
                            if len(ss) == 1:
                                traspasos = traspasos.filter(Q(usuariobienrecibe__nombres__icontains=searchsr, tipo=1) |
                                                             Q(usuariobienrecibe__apellido1__icontains=searchsr,
                                                               tipo=1) |
                                                             Q(usuariobienrecibe__apellido2__icontains=searchsr,
                                                               tipo=1)).order_by('-numero')
                            else:
                                traspasos = traspasos.filter(
                                    Q(usuariobienentrega__apellido1__icontains=ss[0], tipo=1) & Q(
                                        usuariobienentrega__apellido2__icontains=ss[1], tipo=1)).order_by('-numero')
                    elif 'id' in request.GET:
                        ids = request.GET['id']
                        traspasos = TraspasoActivo.objects.filter(id=ids, tipo=1).order_by('-numero')
                    else:
                        traspasos = TraspasoActivo.objects.filter(tipo=1).order_by('-numero')
                    paging = MiPaginador(traspasos, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['searchse'] = searchse if searchse else ""
                    data['searchsr'] = searchsr if searchsr else ""
                    data['ids'] = ids if ids else ""
                    data['traspasos'] = page.object_list
                    data['reporte_0'] = obtener_reporte('acta_entrega_libro')
                    data['reporte_1'] = obtener_reporte('acta_entrega_vehiculo')
                    data['reporte_2'] = obtener_reporte('acta_entrega_otro')
                    return render(request, "inventario_activofijo/actasentrega.html", data)
                except Exception as ex:
                    pass

            elif action == 'tarjeta':
                try:
                    data['title'] = u'Detalle Tarjeta de control'
                    search = None
                    ids = None
                    idt = None
                    if 's' in request.GET:
                        search = request.GET['s']
                        detalle = TarjetaControl.objects.filter(Q(activo__descripcion__icontains=search) |
                                                                Q(activo__codigointerno=search) |
                                                                Q(activo__codigogobierno=search) |
                                                                Q(numero=search), status=True).distinct().order_by(
                            '-numero')
                    elif 'id' in request.GET:
                        ids = request.GET['id']
                        detalle = TarjetaControl.objects.filter(id=ids, status=True).order_by('-numero')
                    else:
                        detalle = TarjetaControl.objects.filter(status=True).order_by('-numero')
                    paging = MiPaginador(detalle, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['idt'] = idt if idt else ""
                    data['detalles'] = page.object_list
                    data['reporte_0'] = obtener_reporte('tarjeta_control')
                    return render(request, "inventario_activofijo/tarjeta.html", data)
                except Exception as ex:
                    pass

            elif action == 'detalletarjeta':
                try:
                    data['title'] = u'Detalle Tarjeta de control'
                    data['tarjeta'] = tarjeta = TarjetaControl.objects.get(pk=int(request.GET['id']))
                    data['detalles'] = tarjeta.detallemantenimiento_set.all()
                    return render(request, "inventario_activofijo/detalletarjeta.html", data)
                except Exception as ex:
                    pass

            elif action == 'confirmacion':
                try:
                    data['title'] = u'Confirmar traslado'
                    data['traslado'] = TrasladoMantenimiento.objects.get(pk=int(request.GET['id']))
                    return render(request, "inventario_activofijo/confirmacion.html", data)
                except:
                    pass

            elif action == 'addarchivobaja':
                try:

                    activo = ActivoFijo.objects.get(pk=request.GET['id'])
                    if activo.archivobaja:
                        data['filtro'] = filtro = ActivoFijo.objects.get(pk=int(request.GET['id']))
                        data['idactivofijo'] = request.GET['id']
                        data['form2'] = ArchivoActivoBajaForm(initial=model_to_dict(filtro))
                    else:
                        form2 = ArchivoActivoBajaForm()
                        data['idactivofijo'] = request.GET['id']
                        data['form2'] = form2

                    template = get_template("inventario_activofijo/addarchivoactivobaja.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'excelmangarantia':
                try:
                    __author__ = 'Unemi'
                    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on',
                                      num_format_str='#,##0.00')
                    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                    title = easyxf(
                        'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    style1 = easyxf(num_format_str='D-MMM-YY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('exp_xls_post_part')
                    ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    response = HttpResponse(content_type="application/ms-excel")
                    fech_ini = request.GET['fechainicio']
                    fech_fin = request.GET['fechafin']
                    tip_act = request.GET['tact']
                    response[
                        'Content-Disposition'] = 'attachment; filename=listadomangarantia' + random.randint(1,
                                                                                                            10000).__str__() + '.xls'

                    columns = [
                        (u"ACTIVO", 15000),
                        (u"PROVEEDOR", 15000),
                        (u"FECHA DE EJECUCION", 3000),
                        (u"TIPO DE ACTIVO", 10000),
                        (u"VALOR", 3000),
                        (u"Hora(s)", 3000),
                        (u"Minuto(s)", 3000),
                        (u"FUNCIONA AL RECIBIR", 8000),
                        (u"FUNCIONA AL ENTREGAR", 8000),
                        (u"USUARIO ENTREGO EQUIPO", 8000),
                        (u"OBSERVACION", 10000),
                        (u"TAREAS DE LIMPIEZA", 10000),
                        (u"DAÑOS ENCONTRADOS", 10000),
                        (u"USUARIO RESP.", 8000)
                    ]
                    row_num = 3
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    date_format = xlwt.XFStyle()
                    date_format.num_format_str = 'yyyy/mm/dd'
                    listamangarantia = MantenimientosActivosGarantia.objects.db_manager('sga_select').filter(
                        status=True, fechainicio__range=[fech_ini, fech_fin], tipoactivo__id=tip_act)
                    row_num = 4
                    for lista in listamangarantia:
                        i = 0
                        campo1 = lista.activofijo.descripcion + ' ' + lista.activofijo.codigointerno + ' ' + lista.activofijo.codigogobierno
                        if lista.proveedor:
                            campo2 = lista.proveedor.nombre
                        else:
                            campo2 = ''
                        campo3 = lista.fechainicio
                        campo5 = lista.valor
                        campo6 = lista.personacreador()
                        campo7 = lista.horamax
                        campo8 = lista.minutomax
                        if lista.estfrec:
                            campo9 = 'SI'
                        else:
                            campo9 = 'NO'
                        if lista.estfent:
                            campo10 = 'SI'
                        else:
                            campo10 = 'NO'
                        if lista.estusu:
                            campo14 = 'SI'
                        else:
                            campo14 = 'NO'
                        campo11 = lista.observacion
                        listatareas = []
                        danioenc = []
                        if lista.tareasactivospreventivosgarantialimp_set.values_list('grupos__descripcion',
                                                                                      flat=True).filter(
                                status=True).exists():
                            campo12 = lista.tareasactivospreventivosgarantialimp_set.values_list('grupos__descripcion',
                                                                                                 flat=True).filter(
                                status=True)
                            for listadocampo12 in campo12:
                                listatareas.append(listadocampo12)
                        else:
                            listatareas.append('NINGUNA')
                        if lista.tareasactivospreventivosgarantiaerr_set.values_list('grupos__descripcion',
                                                                                     flat=True).filter(
                                status=True).exists():
                            campo13 = lista.tareasactivospreventivosgarantiaerr_set.values_list('grupos__descripcion',
                                                                                                flat=True).filter(
                                status=True)
                            for listadocampo13 in campo13:
                                danioenc.append(listadocampo13)
                        else:
                            danioenc.append('NINGUNA')
                        campo15 = lista.tipoactivo.descripcion
                        ws.write(row_num, 0, campo1, font_style2)
                        ws.write(row_num, 1, campo2, font_style2)
                        ws.write(row_num, 2, campo3, date_format)
                        ws.write(row_num, 3, campo15, font_style2)
                        ws.write(row_num, 4, campo5, font_style2)
                        ws.write(row_num, 5, campo7, font_style2)
                        ws.write(row_num, 6, campo8, font_style2)
                        ws.write(row_num, 7, campo9, font_style2)
                        ws.write(row_num, 8, campo10, font_style2)
                        ws.write(row_num, 9, campo14, font_style2)
                        ws.write(row_num, 10, campo11, font_style2)
                        ws.write(row_num, 11, str(listatareas), font_style2)
                        ws.write(row_num, 12, str(danioenc), font_style2)
                        ws.write(row_num, 13, campo6, font_style2)
                        row_num += 1
                    wb.save(response)
                    return response
                except Exception as ex:
                    pass

            elif action == 'excelmanpreventivos':
                try:
                    __author__ = 'Unemi'
                    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on',
                                      num_format_str='#,##0.00')
                    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                    title = easyxf(
                        'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    style1 = easyxf(num_format_str='D-MMM-YY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('exp_xls_post_part')
                    # ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=listadomanpreventivos' + random.randint(1,
                                                                                                                    10000).__str__() + '.xls'
                    fech_ini = request.GET['fechainicio']
                    fech_fin = request.GET['fechafin']
                    tip_act = request.GET['tact']
                    columns = [
                        (u"Activo", 10000),
                        (u"Codigo Interno", 4000),
                        (u"Codigo Gobierno", 4000),
                        (u"Monitor", 4000),
                        (u"Teclado", 4000),
                        (u"Mouse", 4000),
                        (u"Mantenimiento", 4000),
                        (u"Procesador", 4000),
                        (u"Memoria", 4000),
                        (u"Disco duro", 4000),
                        (u"Particiones", 4000),
                        (u"Sistema operativo", 4000),
                        (u"Arquitectura", 4000),
                        (u"Service pack", 4000),
                        (u"Fecha mantenimiento", 4000),
                        (u"Funciona recibe", 4000),
                        (u"Funciona entrega", 4000),
                        (u"Tareas realizadas", 15000),
                        (u"Tareas no realizadas", 15000),
                        (u"Observaciones", 10000),
                        (u"Usuario Ejecuto Mant.", 10000)
                    ]

                    row_num = 0
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    date_format = xlwt.XFStyle()
                    date_format.num_format_str = 'yyyy-mm-dd'
                    date_formatreverse = xlwt.XFStyle()
                    date_formatreverse.num_format_str = 'dd/mm/yyyy'
                    listadomanpreventivos = MantenimientosActivosPreventivos.objects.select_related().filter(
                        fecha__gte=fech_ini, fecha__lte=fech_fin, tipoactivo__id=tip_act, status=True)
                    row_num = 1
                    i = 0
                    for listado in listadomanpreventivos:
                        i += 1
                        campo1 = listado.activofijo.descripcion + ' ' + listado.activofijo.codigointerno + ' ' + listado.activofijo.codigogobierno
                        campo2 = listado.activofijo.codigointerno
                        campo3 = listado.activofijo.codigogobierno
                        campo4 = listado.monitor
                        campo5 = listado.teclado
                        campo6 = listado.mouse
                        campo7 = listado.get_tipomantenimiento_display()
                        campo8 = listado.procesador
                        campo9 = listado.memoria
                        campo10 = listado.discoduro
                        campo11 = listado.particiones
                        campo12 = listado.sistemaoperativo
                        campo13 = listado.arquitectura
                        campo14 = listado.service
                        campo15 = listado.fecha
                        campo16 = 'NO'
                        campo17 = 'NO'
                        if listado.funcionarecibe:
                            campo16 = 'SI'
                        if listado.funcionaentrega:
                            campo17 = 'SI'
                        listatareas = []
                        listatareasno = []
                        if listado.tareasactivospreventivos_set.values_list('grupos__descripcion', flat=True).filter(
                                status=True).exists():
                            campo18 = listado.tareasactivospreventivos_set.values_list('grupos__descripcion',
                                                                                       flat=True).filter(status=True)
                            for listadocampo18 in campo18:
                                listatareas.append(listadocampo18)
                        else:
                            listatareas = ''
                        campo19 = MantenimientoGruCategoria.objects.values_list('descripcion', flat=True).filter(
                            grupocategoria=listado.tipoactivo, status=True).exclude(
                            pk__in=listado.tareasactivospreventivos_set.values_list('grupos__id', flat=True).filter(
                                status=True))
                        for listadocampo19 in campo19:
                            listatareasno.append(listadocampo19)
                        campo20 = listado.observaciones
                        campo21 = listado.personacreador()
                        ws.write(row_num, 0, campo1, font_style2)
                        ws.write(row_num, 1, campo2, font_style2)
                        ws.write(row_num, 2, campo3, font_style2)
                        ws.write(row_num, 3, campo4, font_style2)
                        ws.write(row_num, 4, campo5, font_style2)
                        ws.write(row_num, 5, campo6, font_style2)
                        ws.write(row_num, 6, campo7, font_style2)
                        ws.write(row_num, 7, campo8, font_style2)
                        ws.write(row_num, 8, campo9, font_style2)
                        ws.write(row_num, 9, campo10, font_style2)
                        ws.write(row_num, 10, campo11, font_style2)
                        ws.write(row_num, 11, campo12, font_style2)
                        ws.write(row_num, 12, campo13, font_style2)
                        ws.write(row_num, 13, campo14, font_style2)
                        ws.write(row_num, 14, campo15, date_format)
                        ws.write(row_num, 15, campo16, font_style2)
                        ws.write(row_num, 16, campo17, font_style2)
                        ws.write(row_num, 17, str(listatareas), font_style2)
                        ws.write(row_num, 18, str(listatareasno), font_style2)
                        ws.write(row_num, 19, campo20, font_style2)
                        ws.write(row_num, 20, campo21, font_style2)
                        row_num += 1
                    wb.save(response)
                    return response
                except Exception as ex:
                    pass

            elif action == 'atenciontecnologico':
                try:
                    data['title'] = u'Equipo Tecnológico'
                    search = None
                    ids = None
                    data['codigo'] = codigo = int(request.GET['codigo']) if 'codigo' in request.GET else 0
                    data['baja'] = baja = int(request.GET['baja']) if 'baja' in request.GET else 0
                    data['rangosemaforo'] = RangoVidaUtil.objects.filter(status=True).order_by('anio', 'descripcion')
                    data['grupocatalogo'] = GruposCategoria.objects.filter(status=True)
                    if codigo == 0:
                        activos = ActivoFijo.objects.filter(
                            Q(archivobaja__isnull=True) | Q(archivobaja=''), catalogo__equipoelectronico=True,
                            catalogo__status=True, status=True).order_by('descripcion')
                    else:
                        activos = ActivoFijo.objects.filter(
                            Q(archivobaja__isnull=True) | Q(archivobaja=''), catalogo__equipoelectronico=True,
                            catalogo__status=True, catalogo__grupo__id=codigo, status=True).order_by(
                            'descripcion')
                    if 's' in request.GET:
                        search = request.GET['s'].strip()
                        ss = search.split(' ')
                        if len(ss) == 1:
                            activos = activos.filter(Q(codigogobierno=search) | Q(codigointerno=search),
                                                     status=True)
                    if baja == 1 or baja == 0:
                        activos = activos.filter(statusactivo=1)
                    else:
                        if baja == 2:
                            activos = activos.filter(statusactivo=2)
                    data['totales'] = activos.values('id').count()

                    paging = MiPaginador(activos, 15)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['listadocatalogo'] = page.object_list
                    return render(request, "inventario_activofijo/histatenciontecnologico.html", data)
                except Exception as ex:
                    pass

            elif action == 'configurardirector':
                try:
                    data['title'] = u'Configurar director en archivo de baja'
                    search = None
                    ids = None
                    directores = DirectorResponsableBaja.objects.filter(status=True)
                    if 's' in request.GET:
                        search = request.GET['s'].strip()
                        ss = search.split(' ')
                        if len(ss) == 1:
                            directores = directores.filter(Q(id=search) | Q(responsable__nombres__icontains=search) |
                                                           Q(responsable__apellido1__icontains=search) |
                                                           Q(responsable__apellido2__icontains=search) |
                                                           Q(responsable__cedula__icontains=search)
                                                           , status=True)
                    paging = MiPaginador(directores, 15)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['directores'] = page.object_list
                    return render(request, "inventario_activofijo/viewconfiguradirector.html", data)
                except Exception as ex:
                    pass

            if action == 'adddirector':
                try:
                    data['title'] = u'Adicionar director'
                    data['form2'] = DirectorResponsableBajaForm()
                    template = get_template("inventario_activofijo/modal/formadicionardirector.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass
            if action == 'editdirector':
                try:
                    data['id'] = request.GET['id']
                    data['tipo'] = director = DirectorResponsableBaja.objects.get(pk=request.GET['id'])
                    data['form2'] = DirectorResponsableBaja2Form(initial={'actual': director.actual,
                                                                          'fechainicio': str(director.fechainicio),
                                                                          'fechafin': str(director.fechafin)
                                                                          })
                    template = get_template("inventario_activofijo/modal/formadicionardirector.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'attecpdf':
                try:
                    data = {}
                    data['fechaactual'] = datetime.now()
                    data['activofijo'] = activofijo = ActivoFijo.objects.get(pk=int(request.GET['id']))
                    data['detallemantenimiento'] = HdDetalle_Incidente.objects.filter(incidente__activo=activofijo,
                                                                                      status=True,
                                                                                      incidente__status=True,
                                                                                      resolucion__isnull=False).exclude(
                        resolucion__exact='')
                    data['mantenimientopreventivo'] = activofijo.mantenimientosactivospreventivos_set.filter(
                        status=True)
                    data[
                        'mantenimientogarantia'] = mantenimientogarantia = activofijo.mantenimientosactivosgarantia_set.filter(
                        status=True)
                    data['costomantenimientogarantia'] = \
                    mantenimientogarantia.filter(status=True).aggregate(cantidad=Sum('valor'))['cantidad']
                    return conviert_html_to_pdf('inventario_activofijo/histtecnologico_pdf.html',
                                                {'pagesize': 'A4', 'data': data})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Problemas al ejecutar el reporte. %s" % ex})

            elif action == 'histsingarpdf':
                try:
                    data = {}
                    data['fechaactual'] = datetime.now()
                    data['activofijo'] = activofijo = ActivoFijo.objects.get(pk=int(request.GET['id']))
                    data['mantenimiento'] = mantenimiento = activofijo.mantenimientosactivospreventivos_set.get(
                        status=True, id=int(request.GET['idh']))
                    data['tareasmantenimiento'] = MantenimientoGruCategoria.objects.filter(
                        grupocategoria=mantenimiento.tipoactivo, status=True)
                    data['tareasactivo'] = TareasActivosPreventivos.objects.values_list('grupos_id', flat=True).filter(
                        mantenimiento=mantenimiento, status=True)
                    data['ppmantenimiento'] = HdPiezaPartes.objects.filter(grupocategoria=mantenimiento.tipoactivo,
                                                                           status=True)
                    data['piezaparteactivo'] = PiezaParteActivosPreventivos.objects.filter(mantenimiento=mantenimiento,
                                                                                           status=True)
                    data['daniomantenimiento'] = MantenimientoGruDanios.objects.filter(
                        grupocategoria=mantenimiento.tipoactivo, status=True)
                    data['danioactivo'] = TareasActivosPreventivosDanios.objects.filter(mantenimiento=mantenimiento,
                                                                                        status=True)
                    if mantenimiento.nuevo:
                        return conviert_html_to_pdf('inventario_activofijo/histmantsingar_pdfv2.html',
                                                    {'pagesize': 'A4', 'data': data})
                    else:
                        return conviert_html_to_pdf('inventario_activofijo/histmantsingar_pdf.html',
                                                    {'pagesize': 'A4', 'data': data})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Problemas al ejecutar el reporte. %s" % ex})

            elif action == 'histincidentepdf':
                try:
                    data = {}
                    data['mantenimiento'] = mantenimiento = HdDetalle_Incidente.objects.get(id=request.GET['idh'])
                    return conviert_html_to_pdf('inventario_activofijo/histincidentes_pdf.html',
                                                {'pagesize': 'A4', 'data': data})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Problemas al ejecutar el reporte. %s" % ex})

            elif action == 'histcongarpdf':
                try:
                    data = {}
                    data['fechaactual'] = datetime.now()
                    data['activofijo'] = activofijo = ActivoFijo.objects.get(pk=int(request.GET['id']))
                    data['mantenimiento'] = mantenimiento = activofijo.mantenimientosactivosgarantia_set.get(
                        status=True, id=int(request.GET['idh']))
                    data['tareasmantenimiento'] = MantenimientoGruCategoriaGarantiaLimp.objects.filter(
                        grupocategoria=mantenimiento.tipoactivo, status=True)
                    data['tareasactivo'] = TareasActivosPreventivosGarantiaLimp.objects.values_list('grupos_id',
                                                                                                    flat=True).filter(
                        mantenimiento=mantenimiento, status=True)
                    data['danioenc'] = MantenimientoGruCategoriaGarantiaErr.objects.filter(
                        grupocategoria=mantenimiento.tipoactivo, status=True)
                    data['danios'] = TareasActivosPreventivosGarantiaErr.objects.values_list('grupos_id',
                                                                                             flat=True).filter(
                        mantenimiento=mantenimiento, status=True)
                    return conviert_html_to_pdf('inventario_activofijo/histmantcongar_pdf.html',
                                                {'pagesize': 'A4', 'data': data})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Problemas al ejecutar el reporte. %s" % ex})

            if action == 'addgrupo':
                try:
                    data['form2'] = GrupoBienForm()
                    data['id'] = request.GET['id']
                    data['action'] = request.GET['action']
                    template = get_template("inventario_activofijo/modal/formadicionar.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error mostrar el formulario. %s" % ex})

            elif action == 'subirevidenciaperiodo':
                try:
                    form = EvidenciaPeriodoInventarioTecnologicoForm()
                    data['form2'] = form
                    data['idperiodo'] = request.GET['id']
                    data['action'] = 'subirevidenciaperiodo'
                    template = get_template('inventario_activofijo/modal/formsubirevidenciaperiodo.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'editarevidenciaperiodo':
                try:
                    form = EvidenciaPeriodoInventarioTecnologicoForm()
                    evidencia = EvidenciaPeriodoInventarioTecnologico.objects.get(status=True, id=int(request.GET['id']))
                    form.requerirevidencia()
                    form.fields['nombre'].initial = evidencia.nombre
                    form.fields['descripcion'].initial = evidencia.descripcion
                    form.fields['evidencia'].initial = evidencia.archivo
                    form.renombrar()
                    data['form2'] = form
                    data['idperiodo'] = request.GET['id']
                    data['action'] = 'editarevidenciaperiodo'
                    template = get_template('inventario_activofijo/modal/formsubirevidenciaperiodo.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'eliminarevidenciaperiodoinv':
                try:
                    data['title'] = u'Confirmar eliminación'
                    data['idevidenciaperiodoinv'] = int(request.GET['id'])
                    data['idperiodoinv'] = int(request.GET['idperiodoinv'])
                    return render(request, "inventario_activofijo/eliminarevidenciaperiodoinv.html", data)
                except:
                    pass

            elif action == 'verevidenciaperiodoinv':
                try:
                    data['title'] = u'Evidencias'
                    data['periodoinventariotics'] = PeriodoInventarioAT.objects.get(status=True, id=int(request.GET['idperiodoinv']))
                    # data['listadocatalogo'] = evidenciasperiodo = EvidenciaPeriodoInventarioTecnologico.objects.filter(status=True, periodo_id=int(request.GET['id']))
                    # data['totalevidencias'] = evidenciasperiodo.count()
                    search = None
                    ids = None
                    perfil = None
                    data['idperiodoinv'] = request.GET['idperiodoinv']
                    filtrocatalogo = (Q(status=True) & Q(periodo_id=int(request.GET['idperiodoinv'])))
                    filtro = Q(status=True)
                    if 's' in request.GET:
                        s = request.GET['s']
                        data['s'] = s

                        search = request.GET['s'].strip()
                        ss = search.split(' ')
                        if len(ss) == 1:
                            filtro = filtrocatalogo & ((Q(nombre__icontains=search) |
                                                        Q(descripcion__icontains=search) |
                                                        Q(periodo__nombre__icontains=search)))
                        else:
                            filtro = filtrocatalogo & ((Q(nombre__icontains=ss[0]) &
                                                        Q(nombre__icontains=ss[1])) |
                                                       (Q(descripcion__icontains=ss[0]) &
                                                        Q(descripcion__icontains=ss[1])) |
                                                       (Q(periodo__nombre__icontains=ss[0]) &
                                                        Q(periodo__nombre__icontains=ss[1])))
                    else:
                        filtro = filtrocatalogo
                    evidenciasperiodoinv = EvidenciaPeriodoInventarioTecnologico.objects.filter(filtro).distinct().order_by('-id')
                    data['totales'] = totales = evidenciasperiodoinv.values('id').count()
                    paging = MiPaginador(evidenciasperiodoinv, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['perfil'] = perfil if perfil else ""
                    data['listadocatalogo'] = page.object_list
                    data['totalevidencias'] = totales
                    return render(request, "inventario_activofijo/verevidenciasperiodoinventario.html", data)
                except Exception as ex:
                        pass

            elif action == 'verestadosgenerales':
                try:
                    data['title'] = u'Estados generales'
                    data['idperiodoinv'] = request.GET['idperiodoinv']
                    search = None
                    ids = None
                    perfil = None
                    # data['idperiodoinv'] = request.GET['idperiodoinv']
                    filtrocatalogo = (Q(status=True))
                    filtro = Q(status=True)
                    if 's' in request.GET:
                        s = request.GET['s']
                        data['s'] = s

                        search = request.GET['s'].strip()
                        ss = search.split(' ')
                        if len(ss) == 1:
                            filtro = filtrocatalogo & (Q(descripcion__icontains=search))
                        else:
                            filtro = filtrocatalogo & ((Q(descripcion__icontains=ss[0]) &
                                                        Q(descripcion__icontains=ss[1])))
                    else:
                        filtro = filtrocatalogo
                    estadosgeneralesinv = EstadosGeneralesInventarioAT.objects.filter(filtro).distinct().order_by('-id')
                    data['totales'] = totales = estadosgeneralesinv.values('id').count()
                    paging = MiPaginador(estadosgeneralesinv, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['perfil'] = perfil if perfil else ""
                    data['listadocatalogo'] = page.object_list
                    data['totalevidencias'] = totales
                    return render(request, "inventario_activofijo/verestadosgenerales.html", data)
                except Exception as ex:
                        pass

            elif action == 'adicionarestadogeneral':
                try:
                    form = EstadosGeneralesInventarioATForm()
                    data['action'] = 'adicionarestadogeneral'
                    data['form2'] = form
                    template = get_template('inventario_activofijo/modal/formestadosgenerales.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'eliminarestadogeneral':
                try:
                    data['title'] = u'Confirmar eliminación'
                    data['idestadogeneral'] = int(request.GET['id'])
                    data['idperiodoinv'] = int(request.GET['idperiodoinv'])
                    return render(request, "inventario_activofijo/eliminarestadogeneral.html", data)
                except:
                    pass

            elif action == 'editarestadogeneral':
                try:
                    form = EstadosGeneralesInventarioATForm()
                    estadogeneral = EstadosGeneralesInventarioAT.objects.get(status=True, id=int(request.GET['activo']))
                    form.fields['descripcion'].initial = estadogeneral.descripcion
                    data['form2'] = form
                    data['idestadogeneral'] = estadogeneral.id
                    data['action'] = 'editarestadogeneral'
                    template = get_template('inventario_activofijo/modal/formeditarestadogeneral.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'reporteactivostecnologicos':
                try:
                    __author__ = 'Unemi'
                    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on',
                                      num_format_str='#,##0.00')
                    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                    title = easyxf(
                        'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    style1 = easyxf(num_format_str='D-MMM-YY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('exp_xls_post_part')
                    # ws.write_merge(0, 0, 0, 7, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=listado_activos' + random.randint(1,
                                                                                                                10000).__str__() + '.xls'
                    row_num = 0
                    columns = [
                        (u"N.", 2000),
                        (u"Fecha ingreso", 3000),
                        (u"Responsable", 6000),
                        (u"Fecha asignación", 6000),
                        (u"Catálogo", 6000),
                        (u"Descripción", 6000),
                        (u"Cod. Gob.", 6000),
                        (u"Cod. Int.", 6000),
                        (u"Serie", 6000),
                        (u"Modelo", 6000),
                        (u"Marca", 6000),


                    ]
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    date_format = xlwt.XFStyle()
                    date_format.num_format_str = 'yyyy/mm/dd'
                    eventoperiodo = int(request.GET['id'])
                    resultado = None
                    if 'tipoperfilpersona' in request.GET:
                        tipoperfilpersona = int(request.GET['tipoperfilpersona'])
                        #Solo administrativos
                        if tipoperfilpersona == 1:
                            resultado = Persona.objects.filter(Q(status=True) & (Q(perfilusuario__administrativo__isnull=False) & Q(
                                perfilusuario__profesor__isnull=True))).distinct()
                        #Solo docentes
                        if tipoperfilpersona == 2:
                            resultado = Persona.objects.filter(Q(status=True) & (
                                (Q(perfilusuario__administrativo__isnull=False) & Q(perfilusuario__profesor__isnull=False))) |
                                (Q(perfilusuario__administrativo__isnull=True) & Q(perfilusuario__profesor__isnull=False))
                                ).distinct()
                        #Ambos
                        if tipoperfilpersona == 0:
                            resultado = Persona.objects.filter(Q(status=True) & (Q(perfilusuario__administrativo__isnull=False) | Q(
                                perfilusuario__profesor__isnull=False))).distinct()
                    total_personas = resultado

                    row_num = 0
                    for lista in total_personas:
                        activos_asignado = ActivoTecnologico.objects.filter(status=True, activotecnologico__status=True, activotecnologico__catalogo__status=True,
                                                                           activotecnologico__catalogo__clasificado=True, activotecnologico__catalogo__equipoelectronico=True,
                                                                           activotecnologico__statusactivo=1,
                                                                            activotecnologico__responsable=lista)
                        if activos_asignado:
                            for activo in activos_asignado:
                                fecha_traspaso = None
                                fecha_asignacion = activo.activotecnologico.detalletraspasoactivo_set.filter(status=True).order_by('-codigotraspaso__fecha')
                                if fecha_asignacion:
                                    fecha_traspaso = fecha_asignacion[0].codigotraspaso.fecha
                                row_num += 1
                                ws.write(row_num, 0, row_num, font_style2)
                                ws.write(row_num, 1, activo.activotecnologico.fechaingreso, font_style2)
                                ws.write(row_num, 2, activo.activotecnologico.responsable.__str__(), font_style2)
                                ws.write(row_num, 3, fecha_traspaso, font_style2)
                                ws.write(row_num, 4, activo.activotecnologico.catalogo.__str__(), font_style2)
                                ws.write(row_num, 5, activo.activotecnologico.descripcion, font_style2)
                                ws.write(row_num, 6, activo.activotecnologico.codigogobierno, font_style2)
                                ws.write(row_num, 7, activo.activotecnologico.codigointerno, font_style2)
                                ws.write(row_num, 8, activo.activotecnologico.serie, font_style2)
                                ws.write(row_num, 9, activo.activotecnologico.modelo, font_style2)
                                ws.write(row_num, 10, activo.activotecnologico.marca, font_style2)



                        # ws.write(row_num, 7, campo7, font_style2)
                    wb.save(response)
                    return response
                except Exception as ex:
                    pass

            #Cronnograma
            elif action == 'cronogramav1':
                try:
                    data['title'] = u'Cronograma'
                    data['idp'] = idp = int(encrypt(request.GET['id']))
                    data['periodo']=PeriodoInventarioAT.objects.get(id=idp)
                    fecha=hoy.date()
                    data['hoy_str']=fecha
                    return render(request, 'inventario_activofijo/cronograma.html', data)
                except Exception as ex:
                    return render({'result': False, 'mensaje': '{}'.format(ex)})

            elif action == 'detalleconstataciones':
                try:
                    #url_vars = f"&action=detalleconstataciones"
                    data['title'] = u'Cronograma'
                    data['idp'] = idp = int(encrypt(request.GET['idp']))
                    data['periodo_c'] = periodo_c = PeriodoInventarioAT.objects.get(id=idp)
                    search = request.GET.get('s', '')
                    desde = request.GET.get('desde', '')
                    hasta = request.GET.get('hasta', '')
                    estado = request.GET.get('estado', '')
                    constatador = request.GET.get('constatador', None)
                    filtro =  Q(status=True,periodo_id=idp)
                    url_vars = f"&action={action}&idp={request.GET['idp']}"
                    if search:
                        q = search.upper().strip()
                        s = q.split(" ")
                        if len(s) == 1:
                            filtro = filtro & (Q(persona__nombres__icontains=q) |
                                               Q(persona__apellido1__icontains=q) |
                                               Q(persona__cedula__icontains=q) |
                                               Q(persona__apellido2__icontains=q) |
                                               Q(persona__cedula__contains=q))
                        elif len(s) == 2:
                            filtro = filtro & ((Q(persona__apellido1__contains=s[0]) & Q(persona__apellido2__contains=s[1])) |
                                               (Q(persona__nombres__icontains=s[0]) & Q(persona__nombres__icontains=s[1])) |
                                               (Q(persona__nombres__icontains=s[0]) & Q(persona__apellido1__contains=s[1])))
                        else:
                            filtro = filtro & ((Q(persona__nombres__contains=s[0]) & Q(persona__apellido1__contains=s[1]) & Q(persona__apellido2__contains=s[2])) |
                                               (Q(persona__nombres__contains=s[0]) & Q(persona__nombres__contains=s[1]) & Q(persona__apellido1__contains=s[2])))
                        url_vars += f"&s={search}"
                        data['s'] = search
                    if desde:
                        filtro = filtro & (Q(fecha__gte=desde))
                        url_vars += f"&desde={desde}"
                        data['desde'] = desde
                    if hasta:
                        filtro = filtro & (Q(fecha__lte=hasta))
                        url_vars += f"&hasta={hasta}"
                        data['hasta'] = hasta
                    if estado:
                        filtro = filtro & (Q(estado=estado))
                        url_vars += f"&estado={estado}"
                        data['estado'] = int(estado)
                    if constatador:
                        filtro = filtro & (Q(responsable_id=constatador))
                        url_vars += f"&constatador={constatador}"
                        data['constatador'] = int(constatador)
                    cronogramas = CronogramaPersonaConstatacionAT.objects.filter(filtro)
                    if 'exportar_excel' in request.GET:
                        wb = openxl.Workbook()
                        wb["Sheet"].title = "Reporte_activos_tecnologicos"
                        ws = wb.active
                        style_title = openxlFont(name='Arial', size=16, bold=True)
                        style_cab = openxlFont(name='Arial', size=10, bold=True)
                        alinear = alin(horizontal="center", vertical="center")
                        thin = Side(border_style="thin", color="000000")
                        response = HttpResponse(content_type="application/ms-excel")
                        response['Content-Disposition'] = 'attachment; filename="Reporte de constataciones de activos tecnológicos' + '-' + random.randint(1, 10000).__str__() + '.xlsx"'
                        ws.column_dimensions['B'].width = 25
                        ws.column_dimensions['C'].width = 35
                        ws.column_dimensions['D'].width = 15
                        ws.column_dimensions['E'].width = 15
                        ws.column_dimensions['F'].width = 25
                        ws.column_dimensions['G'].width = 15
                        ws.column_dimensions['H'].width = 15
                        ws.column_dimensions['I'].width = 15
                        ws.column_dimensions['J'].width = 40
                        ws.merge_cells('A1:J1')
                        ws['A1'] = 'CONSTATACIONES'
                        celda1 = ws['A1']
                        celda1.font = style_title
                        celda1.alignment = alinear
                        celdaconst = ws['A1:J1']
                        for row in celdaconst:
                            for cell in row:
                                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                        ws.merge_cells('B2:F2')
                        ws['B2'] = 'FUNCIONARIO A CARGO'
                        celda2 = ws['B2']
                        celda2.font = style_cab
                        celda2.alignment = alinear
                        celdafunc = ws['B2:F2']
                        for row in celdafunc:
                            for cell in row:
                                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                        ws.merge_cells('G2:J2')
                        ws['G2'] = 'CONSTATACIÓN'
                        celda3 = ws['G2']
                        celda3.font = style_cab
                        celda3.alignment = alinear
                        celdacon = ws['G2:J2']
                        for row in celdacon:
                            for cell in row:
                                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                        columns = [u"N°", u"NOMBRE", u"CARGO", u"IDENTIFICACION", u"TELEFONO",
                                   u"EMAIL", u"ESTADO", u"FECHA", u"HORA", u"DESCRIPCION"
                                   ]
                        row_num = 3
                        for col_num in range(0, len(columns)):
                            celda = ws.cell(row=row_num, column=(col_num + 1), value=columns[col_num])
                            celda.font = style_cab
                            celda.alignment = alinear
                            celda.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                        row_num = 4
                        numero = 1
                        for list in cronogramas:
                            numerocelda = ws.cell(row=row_num, column=1, value=numero)
                            numerocelda.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                            nombrecelda = ws.cell(row=row_num, column=2, value=str(list.persona.nombre_completo_minus()))
                            nombrecelda.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                            cargocelda = ws.cell(row=row_num, column=3, value=str(list.persona.cargo_persona().denominacionpuesto.descripcion) if list.persona.cargo_persona() else 'SIN CARGO ASIGNADO')
                            cargocelda.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                            cedulacelda = ws.cell(row=row_num, column=4, value=str(list.persona.cedula))
                            cedulacelda.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                            telefonocelda = ws.cell(row=row_num, column=5, value=str(list.persona.telefono))
                            telefonocelda.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                            emailcelda = ws.cell(row=row_num, column=6, value=str(list.persona.emailinst))
                            emailcelda.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                            estadocelda = ws.cell(row=row_num, column=7, value=str(list.get_estado_display()))
                            estadocelda.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                            fechacelda = ws.cell(row=row_num, column=8, value=str(list.fecha))
                            fechacelda.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                            horacelda = ws.cell(row=row_num, column=9, value=str(list.hora))
                            horacelda.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                            desccelda = ws.cell(row=row_num, column=10, value=str(list.descripcion))
                            desccelda.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                            row_num += 1
                            numero += 1
                        wb.save(response)
                        return response
                    if 'exportar_excel_constataciones' in request.GET:
                        namefile='Reporte_activos_constatados'
                        pagina='Activos constatados y no constatados'
                        wb = openxl.Workbook()
                        wb["Sheet"].title = namefile
                        ws = wb.active
                        style_title = openxlFont(name='Arial', size=16, bold=True)
                        style_cab = openxlFont(name='Arial', size=10, bold=True)
                        alinear = alin(horizontal="center", vertical="center")
                        thin = Side(border_style="thin", color="000000")
                        response = HttpResponse(content_type="application/ms-excel")
                        response['Content-Disposition'] = f'attachment; filename="{namefile}' + '-' + random.randint(1, 10000).__str__() + '.xlsx"'
                        ws.column_dimensions['B'].width = 30
                        ws.column_dimensions['C'].width = 15
                        ws.column_dimensions['D'].width = 15
                        ws.column_dimensions['E'].width = 25
                        ws.column_dimensions['F'].width = 15
                        ws.column_dimensions['G'].width = 15
                        ws.column_dimensions['H'].width = 20
                        ws.column_dimensions['I'].width = 15
                        ws.column_dimensions['J'].width = 15
                        ws.column_dimensions['L'].width = 40
                        ws.merge_cells('A1:L1')
                        ws['A1'] = pagina
                        celda1 = ws['A1']
                        celda1.font = style_title
                        celda1.alignment = alinear
                        celdaconst = ws['A1:L1']
                        for row in celdaconst:
                            for cell in row:
                                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                        columns = [u"N°", u"Nombre", u"Codigo Gobierno","Codigo Interno", u"Cargo",
                                   u"Identificación", u"Teléfono", u"Email", u"Estado", u"Fecha", u"Hora", u"Descripción",u"Estado activo",
                                   u"Constatado",u"Proceso baja", u"Encontrado",u"En uso", u"Pertenece a usuario",u"Requiere traspaso", u"Quien usa"
                                   ]
                        row_num = 2
                        for col_num in range(0, len(columns)):
                            # if not col_num == 3:
                            #     celda=ws.merge_cells(start_row=2, end_row=3, start_column=(col_num + 1), end_column=(col_num + 1))
                            celda = ws.cell(row=row_num, column=(col_num + 1), value=columns[col_num])
                            celda.font = style_cab
                            celda.alignment = alinear
                            celda.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                        row_num = 3
                        numero = 1
                        for list in cronogramas:
                            for constatacion in list.constataciones():
                                numerocelda = ws.cell(row=row_num, column=1, value=numero)
                                nombrecelda = ws.cell(row=row_num, column=2, value=str(list.persona.nombre_completo_minus()))
                                codigogob = ws.cell(row=row_num, column=3, value=str(constatacion.activo.activotecnologico.codigogobierno) if constatacion.activo.activotecnologico.codigogobierno else '')
                                codigointerno = ws.cell(row=row_num, column=4, value=str(constatacion.activo.activotecnologico.codigointerno) if constatacion.activo.activotecnologico.codigointerno else '')
                                cargocelda = ws.cell(row=row_num, column=5,value=str(list.persona.cargo_persona().denominacionpuesto.descripcion) if list.persona.cargo_persona() else 'SIN CARGO ASIGNADO')
                                identificacion = ws.cell(row=row_num, column=6, value=str(list.persona.cedula))
                                telefonocelda = ws.cell(row=row_num, column=7, value=str(list.persona.telefono))
                                emailcelda = ws.cell(row=row_num, column=8, value=str(list.persona.emailinst))
                                estadocelda = ws.cell(row=row_num, column=9, value=str(list.get_estado_display()))
                                fechacelda = ws.cell(row=row_num, column=10, value=str(list.fecha))
                                horacelda = ws.cell(row=row_num, column=11, value=str(list.hora))
                                desccelda = ws.cell(row=row_num, column=12, value=str(list.descripcion))
                                estado_activo = ws.cell(row=row_num, column=13, value=str(constatacion.estadoactual))
                                constatado = ws.cell(row=row_num, column=14, value=str('SI' if constatacion.constatado else '' ))
                                procesobaja = ws.cell(row=row_num, column=15, value=str('SI' if constatacion.activo.activotecnologico.procesobaja else '' ))
                                encontrado = ws.cell(row=row_num, column=16, value=str('SI' if constatacion.encontrado else ''))
                                enuso = ws.cell(row=row_num, column=17, value=str('SI' if constatacion.enuso else ''))
                                perteneceusuario = ws.cell(row=row_num, column=18, value=str('SI' if constatacion.perteneceusuario else ''))
                                traspaso = ws.cell(row=row_num, column=19, value=str('SI' if constatacion.requieretraspaso else ''))
                                usuariobien = ws.cell(row=row_num, column=20, value=str(constatacion.usuariobienes if constatacion.usuariobienes else '' ))
                                row_num += 1
                                numero += 1
                        wb.save(response)
                        return response
                    paging = MiPaginador(cronogramas, 10)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data["url_vars"] = url_vars
                    data['constataciones'] = page.object_list
                    data['constatadores']=CronogramaPersonaConstatacionAT.objects.filter(status=True,periodo_id=idp,responsable__isnull=False).values_list('responsable','responsable__nombres','responsable__apellido1','responsable__apellido2').distinct()
                    data['estados'] = ESTADO_CONSTATACION_AT
                    data['reporte_0'] = obtener_reporte('acta_constatacion_activo_tecnologico')
                    # CONTADOR
                    data['contplanificados'] = len(cronogramas.filter(estado=1))
                    data['contenproceso'] = len(cronogramas.filter(estado=2))
                    data['contfinalizados'] = len(cronogramas.filter(estado=3))
                    data['contcerrados'] = len(cronogramas.filter(estado=4))
                    total_activos, total_constatados, total_por_constatar = 0, 0, 0
                    for c in cronogramas:
                        total_activos += len(c.activos_asignados())
                        total_constatados += len(c.activos_constatados())
                    data['total_activos'] = total_activos
                    data['total_constatados'] = total_constatados
                    data['total_por_constatar'] = total_activos - total_constatados
                    return render(request, 'inventario_activofijo/detalleconstataciones.html', data)
                except Exception as ex:
                    return render({'result': False, 'mensaje': '{}'.format(ex)})

            elif action == 'verdetalle_constataciones':
                try:
                    id = int(encrypt(request.GET['id']))
                    data['title'] = 'Detalle de constatación'
                    data['cronograma'] = CronogramaPersonaConstatacionAT.objects.get(id=id)
                    data['activosconstatados'] = DetalleConstatacionFisicaActivoTecnologico.objects.filter(Q(status=True, cronograma_id=id, constatado=True))
                    data['activosnoconstatados'] = DetalleConstatacionFisicaActivoTecnologico.objects.filter(Q(status=True, cronograma_id=id, constatado=False))
                    template = get_template("inventario_activofijo/modal/verdetalle_constataciones.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})


            elif action == 'constatacionesresponsables':
                try:
                    data['idp'] = idp = int(encrypt(request.GET['id']))
                    base_query = Q(status=True, periodo_id=idp,  detalleconstatacionfisicaactivotecnologico__constatado = True, responsable_id__isnull=False,
                                                                                detalleconstatacionfisicaactivotecnologico__isnull=False,
                                                                                detalleconstatacionfisicaactivotecnologico__status=True)
                    filtro = base_query
                    data['title'] = u'Constataciones por responsable'
                    desde = request.GET.get('desde', '')
                    hasta = request.GET.get('hasta', '')
                    after = request.GET.get('after', '')
                    before = request.GET.get('before', '')
                    url_vars = f"&action={action}"
                    if after:
                        filtro = filtro & Q(fecha__gt=after)
                    if before:
                        filtro = filtro & Q(fecha__lt=before)
                        fechas = CronogramaPersonaConstatacionAT.objects.filter(filtro).values_list('fecha', flat=True).distinct().order_by('-fecha')
                    else:
                        fechas = CronogramaPersonaConstatacionAT.objects.filter(filtro).values_list('fecha', flat=True).distinct().order_by('fecha')

                    if desde and hasta:
                        filtro = filtro & (Q(fecha__range=[desde, hasta]))
                        lenafter = 0
                        if not url_vars == '':
                            url_vars += f"&desde={desde}&hasta={hasta}"
                            data['url_vars'] = url_vars
                    else:
                        lenfecha = fechas.__len__()
                        lenafter = lenfecha
                        lenbefore = lenfecha
                        fec = 7
                        if fechas.__len__() > 0:
                            if lenfecha >= fec:
                                fechas = fechas[:fec]
                                lenfecha = fechas.__len__()
                            if after:
                                lenafter = lenafter - fec
                                lenbefore = lenbefore + fec
                            if before:
                                lenbefore = lenbefore - fec
                                lenafter = lenafter + fec
                                fechas.reverse()
                            if 'desde' not in request.GET and 'hasta' not in request.GET:
                                desde = fechas[0].strftime('%Y-%m-%d')  # Formato YYYY-MM-DD
                                hasta = fechas[lenfecha - 1].strftime('%Y-%m-%d')  # Formato DD/MM/YYYY
                                filtro = filtro & (Q(fecha__range=[desde, hasta]))
                                url_vars = ''
                    if 'after' not in request.GET and 'before' not in request.GET:
                        lenbefore = 0

                    data['lenafter'] = lenafter
                    data['lenbefore'] = lenbefore
                    data['desde'] = desde
                    data['hasta'] = hasta
                    data['fechas'] = fechas = CronogramaPersonaConstatacionAT.objects.filter(filtro).values_list('fecha', flat=True).distinct().order_by('fecha')
                    idresp = CronogramaPersonaConstatacionAT.objects.filter(status=True).values_list('responsable_id', flat=True).order_by('responsable_id')
                    data['respons'] = respons = Persona.objects.filter(id__in=idresp)
                    data['constatadores'] = respons.values_list('id', 'nombres', 'apellido1', 'apellido2').order_by('id').distinct()
                    data['totalporfecha'] = []
                    for fecha in fechas:
                        t = DetalleConstatacionFisicaActivoTecnologico.objects.filter(cronograma__status=True, status=True, constatado= True, cronograma__fecha=fecha, cronograma__responsable_id__in=respons).count()
                        data['totalporfecha'].append(t)

                    return render(request, 'inventario_activofijo/constatacionesresponsables.html', data)
                except Exception as ex:
                    pass

            elif action == 'cargareventos':
                try:
                    idp = int(encrypt(request.GET['id']))
                    fecha=hoy.date()
                    filtro=Q(status=True,periodo_id=idp)
                    value=request.GET.get('value','')
                    if value:
                        q = request.GET['value'].upper().strip()
                        s = q.split(" ")
                        if len(s) == 1:
                            filtro = filtro &(Q(persona__nombres__icontains=q) |
                                               Q(persona__apellido1__icontains=q) |
                                               Q(persona__cedula__icontains=q) |
                                               Q(persona__apellido2__icontains=q) |
                                               Q(persona__cedula__contains=q))
                        elif len(s) == 2:
                            filtro = filtro & ((Q(persona__apellido1__contains=s[0]) & Q(persona__apellido2__contains=s[1])) |
                                                (Q(persona__nombres__icontains=s[0]) & Q(persona__nombres__icontains=s[1])) |
                                                 (Q(persona__nombres__icontains=s[0]) & Q(persona__apellido1__contains=s[1])))
                        else:
                            filtro = filtro & ((Q(persona__nombres__contains=s[0]) & Q(persona__apellido1__contains=s[1]) & Q(persona__apellido2__contains=s[2])) |
                                                (Q(persona__nombres__contains=s[0]) & Q(persona__nombres__contains=s[1]) & Q(persona__apellido1__contains=s[2])))
                    else:
                        finicio=datetime.strptime(request.GET['finicio'], '%d/%m/%Y')
                        ffin=datetime.strptime(request.GET['ffin'], '%d/%m/%Y')
                        filtro = filtro & Q(fecha__range=(finicio,ffin))
                    cronogramas=CronogramaPersonaConstatacionAT.objects.filter(filtro)
                    if cronogramas and value:
                        fecha = cronogramas.order_by('fecha').first().fecha
                    event_list = []
                    for cronograma in cronogramas:
                        start_date = datetime.combine(cronograma.fecha, cronograma.hora).strftime('%Y-%m-%dT%H:%M:%S')
                        end_date = datetime.combine(cronograma.fecha, cronograma.hora).strftime('%Y-%m-%dT%H:%M:%S')
                        event_list.append({
                            'id': cronograma.id,
                            'title':cronograma.persona.nombre_normal_minus(),
                            'extendedProps': {'description': cronograma.descripcion.capitalize() if cronograma.descripcion else 'Sin descripción',
                                              'estado':cronograma.get_estado_display(),
                                              'id_estado':cronograma.estado,
                                              'puede_eliminar':cronograma.puede_eliminar(),
                                              'color_estado':cronograma.color_estado(),
                                              'fecha_c': str(cronograma.fechainicio) if cronograma.fechainicio else '',
                                              'eventColor': cronograma.color_evento()},
                            'start': start_date,
                            'end': end_date,
                            'backgroundColor':cronograma.color_evento(),
                        })

                    return JsonResponse({"fecha_e": str(fecha),'eventos':event_list}, safe=False)
                except Exception as ex:
                    return render({'result': False, 'mensaje': '{}'.format(ex)})

            elif action == 'addevento':
                try:
                    data['idp']=request.GET['idp']
                    data['fecha'] = fecha = request.GET.get('a_extra','')
                    form = CronogramaPersonaInventarioForm()
                    form.fields['persona'].queryset=Persona.objects.none()
                    data['form'] = form
                    template = get_template("inventario_activofijo/modal/formeventos.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'editevento':
                try:
                    data['id']=id=request.GET['id']
                    cronograma=CronogramaPersonaConstatacionAT.objects.get(id=id)
                    form = CronogramaPersonaInventarioForm(initial=model_to_dict(cronograma))
                    form.fields['persona'].queryset=Persona.objects.filter(id=cronograma.persona.id)
                    data['form'] = form
                    template = get_template("inventario_activofijo/modal/formeventos.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'buscarpersonasactivos':
                try:
                    q = request.GET['q'].upper().strip()
                    s = q.split(" ")
                    idp=int(encrypt(request.GET['idp']))
                    idsexcluidas= CronogramaPersonaConstatacionAT.objects.filter(periodo_id=idp, status=True).values_list('persona_id', flat=True)
                    informe_ab = InformeActivoBaja.objects.filter(status=True).values_list('activofijo_id', flat=True)
                    # idspersonas=ActivoFijo.objects.filter(status=True, procesobaja=False, statusactivo=1).exclude(id__in=informe_ab).values_list('responsable_id', flat=True).distinct()
                    idspersonas = ActivoTecnologico.objects.filter(status=True,
                                                                   activotecnologico__statusactivo=1,
                                                                   activotecnologico__procesobaja=False).exclude(activotecnologico_id__in=informe_ab).values_list('activotecnologico__responsable_id', flat=True).distinct()
                    qspersona = Persona.objects.filter(status=True, distributivopersona__isnull=False, id__in=idspersonas).exclude(id__in=idsexcluidas).order_by('apellido1')
                    if len(s) == 1:
                        qspersona = qspersona.filter((Q(nombres__icontains=q) | Q(apellido1__icontains=q) | Q(
                            cedula__icontains=q) | Q(apellido2__icontains=q) | Q(cedula__contains=q)),
                                                     Q(status=True)).distinct()[:15]
                    elif len(s) == 2:
                        qspersona = qspersona.filter((Q(apellido1__contains=s[0]) & Q(apellido2__contains=s[1])) |
                                                     (Q(nombres__icontains=s[0]) & Q(nombres__icontains=s[1])) |
                                                     (Q(nombres__icontains=s[0]) & Q(apellido1__contains=s[1]))).filter(
                            status=True).distinct()[:15]
                    else:
                        qspersona = qspersona.filter(
                            (Q(nombres__contains=s[0]) & Q(apellido1__contains=s[1]) & Q(apellido2__contains=s[2])) |
                            (Q(nombres__contains=s[0]) & Q(nombres__contains=s[1]) & Q(
                                apellido1__contains=s[2]))).filter(status=True).distinct()[:15]
                    resp = [{'id': qs.pk, 'text': f"{qs.nombre_completo_inverso()}",
                             'documento': qs.documento(),
                             'foto': qs.get_foto()} for qs in qspersona]
                    return HttpResponse(json.dumps({'status': True, 'results': resp}))
                except Exception as ex:
                    pass

            elif action == 'buscarpersonas':
                try:
                    q = request.GET['q'].upper().strip()
                    s = q.split(" ")
                    qspersona = Persona.objects.filter((Q(distributivopersona__isnull=False) | Q(administrativo__isnull=False))).order_by('apellido1')
                    if len(s) == 1:
                        qspersona = qspersona.filter((Q(nombres__icontains=q) | Q(apellido1__icontains=q) | Q(
                            cedula__icontains=q) | Q(apellido2__icontains=q) | Q(cedula__contains=q)),
                                                     Q(status=True)).distinct()[:15]
                    elif len(s) == 2:
                        qspersona = qspersona.filter((Q(apellido1__contains=s[0]) & Q(apellido2__contains=s[1])) |
                                                     (Q(nombres__icontains=s[0]) & Q(nombres__icontains=s[1])) |
                                                     (Q(nombres__icontains=s[0]) & Q(apellido1__contains=s[1]))).filter(
                            status=True).distinct()[:15]
                    else:
                        qspersona = qspersona.filter(
                            (Q(nombres__contains=s[0]) & Q(apellido1__contains=s[1]) & Q(apellido2__contains=s[2])) |
                            (Q(nombres__contains=s[0]) & Q(nombres__contains=s[1]) & Q(
                                apellido1__contains=s[2]))).filter(status=True).distinct()[:15]
                    resp = [{'id': qs.pk, 'text': f"{qs.nombre_completo_inverso()}",
                             'documento': qs.documento(),
                             'foto': qs.get_foto()} for qs in qspersona]
                    return HttpResponse(json.dumps({'status': True, 'results': resp}))
                except Exception as ex:
                    pass

            #Constataciones
            elif action == 'constatacionesat':
                try:
                    data['title'] = u'Constataciones'
                    data['id']=id=request.GET.get('id','0')
                    data['list']=request.GET.get('list','')
                    data['cronograma']=cronograma=CronogramaPersonaConstatacionAT.objects.get(id=id)
                    estado, search, filtro, url_vars = int(request.GET.get('estado', '0')),request.GET.get('s', ''), Q(status=True,cronograma_id=cronograma.id), f'&action={action}&id={id}'
                    if search:
                        filtro = filtro & (Q(activo__descripcion__unaccent__icontains=search) | Q(activo__activotecnologico__codigogobierno__icontains=search) | Q(activo__activotecnologico__codigointerno__icontains=search) | Q(activo__codigotic__icontains=search))
                        url_vars += '&s=' + search
                        data['s'] = search

                    if estado != 0:
                        url_vars += f'&estado={estado}'
                        data['estado'] = estado
                        constatado=True if estado==1 else False
                        filtro = filtro & Q(constatado=constatado)

                    listado = DetalleConstatacionFisicaActivoTecnologico.objects.filter(filtro)
                    paging = MiPaginador(listado, 10)
                    p = 1

                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['hoy']=hoy.date()
                    data["url_vars"] = url_vars
                    data['listado'] = page.object_list
                    data['t_constatados'] = t_constatados=len(cronograma.constataciones().filter(constatado=True))
                    data['t_activos'] = t_activos=len(cronograma.constataciones())
                    data['t_porconstatar'] = t_activos-t_constatados
                    if 'exportar_excel' in request.GET:
                        wb = openxl.Workbook()
                        wb["Sheet"].title = "Reporte_constatación "
                        ws = wb.active
                        style_title = openxlFont(name='Arial', size=16, bold=True)
                        style_cab = openxlFont(name='Arial', size=10, bold=True)
                        alinear = alin(horizontal="center", vertical="center")
                        response = HttpResponse(content_type="application/ms-excel")
                        response[
                            'Content-Disposition'] = 'attachment; filename=Reporte de citas' + '-' + random.randint(
                            1, 10000).__str__() + '.xlsx'
                        ws.column_dimensions['B'].width = 25
                        ws.column_dimensions['C'].width = 15
                        ws.column_dimensions['D'].width = 20
                        ws.column_dimensions['E'].width = 25
                        ws.merge_cells('A1:F1')
                        ws['A1'] = 'REPORTE DE ACTIVOS CONSTATADOS'
                        celda1 = ws['A1']
                        celda1.font = style_title
                        celda1.alignment = alinear

                        ws.merge_cells('A2:B2')
                        ws['A2'] = 'Funcionario a cargo:'
                        celda2 = ws['A2']
                        celda2.font = style_cab
                        ws.merge_cells('C2:D2')
                        ws['C2'] = cronograma.persona.nombre_completo_minus()
                        celda3 = ws['C2']

                        columns = [u"N°", u"C.GOBIERNO",u"C.INTERNO", u"C.TICS", u"ACTIVO TECNOLÓGICO",
                                   u"CONSTATADO"
                                   ]
                        row_num = 3
                        for col_num in range(0, len(columns)):
                            celda = ws.cell(row=row_num, column=(col_num + 1), value=columns[col_num])
                            celda.font = style_cab

                        mensaje = 'NO REGISTRA'
                        row_num = 4
                        numero = 0
                        for list in listado:
                            numero += 1
                            constatado = 'NO'
                            if list.constatado:
                                constatado = 'SI'

                            ws.cell(row=row_num, column=1, value=numero)
                            ws.cell(row=row_num, column=2, value=str(list.activo.activotecnologico.codigogobierno))
                            ws.cell(row=row_num, column=3, value=str(list.activo.activotecnologico.codigointerno))
                            ws.cell(row=row_num, column=4, value=str(list.activo.codigotic))
                            ws.cell(row=row_num, column=5, value=str(list.activo))
                            ws.cell(row=row_num, column=6, value=constatado)
                            row_num += 1
                        wb.save(response)
                        return response

                    return render(request, 'inventario_activofijo/constataciones.html', data)
                except Exception as ex:
                    return render({'result': False, 'mensaje': '{}'.format(ex)})

            elif action == 'addconstatacionat':
                try:
                    data['idcronograma']=request.GET['idp']
                    data['activo']=activo=ActivoTecnologico.objects.get(id=encrypt(request.GET['id']))
                    data['componenetes']=ComponenteCatalogoActivo.objects.filter(status=True,catalogo=activo.activotecnologico.catalogo.id)
                    data['estados']=EstadoProducto.objects.filter(status=True)
                    form = ConstatacionFisicaATForm()
                    form.fields['ubicacion'].queryset=HdBloqueUbicacion.objects.none()
                    form.fields['usuariobienes'].queryset=Persona.objects.none()
                    data['form'] = form
                    template = get_template("inventario_activofijo/modal/formconstataciones.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'editconstatacionat':
                try:
                    data['id'] = id = int(encrypt(request.GET['id']))
                    constatacion=DetalleConstatacionFisicaActivoTecnologico.objects.get(id=id)
                    idpersona=constatacion.usuariobienes.id if constatacion.usuariobienes else None
                    form = ConstatacionFisicaATForm(initial=model_to_dict(constatacion))
                    form.fields['usuariobienes'].queryset=Persona.objects.filter(id=idpersona)
                    data['componenetes']=ComponenteCatalogoActivo.objects.filter(status=True,catalogo=constatacion.activo.activotecnologico.catalogo.id)
                    form.fields['ubicacion'].queryset = HdBloqueUbicacion.objects.none()
                    if constatacion.bloque:
                        form.fields['ubicacion'].queryset = HdBloqueUbicacion.objects.filter(status=True,bloque=constatacion.bloque)
                    data['activo'] = constatacion.activo
                    data['aplica']=constatacion.aplica
                    data['estados']=EstadoProducto.objects.filter(status=True)
                    data['form'] = form
                    template = get_template("inventario_activofijo/modal/formconstataciones.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'detalle_activo':
                try:
                    data['activo'] = activo = ActivoFijo.objects.get(pk=int(encrypt(request.GET['id'])))
                    template = get_template("at_activostecnologicos/detalleActM.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, "mensaje": u"Error al obtener los datos."})

            elif action == 'selectubicacion':
                try:
                    if 'id' in request.GET:
                        lista = []
                        ubicaciones = HdBloqueUbicacion.objects.filter(bloque_id=int(request.GET['id']),status=True)
                        for ubi in ubicaciones:
                            lista.append([ubi.id, ubi.ubicacion.nombre])
                        data = {"results": "ok", 'lista': lista}
                        return JsonResponse(data)
                except Exception as ex:
                    pass

            elif action == 'firmaracta':
                try:
                    cronograma_id=request.GET['id']
                    reporte_id = request.GET['reporte']
                    reporte = Reporte.objects.get(pk=reporte_id)
                    base_url = request.META['HTTP_HOST']
                    d = datetime.now()
                    pdfname = reporte.nombre + d.strftime('%Y%m%d_%H%M%S')
                    tipo = 'pdf'
                    paRequest = {
                            'id': cronograma_id,
                            'imp_logo':True,
                            'imp_encabezado':True,
                            'imp_fecha':True,
                            'imp_membretada':False,
                            'url_qr':unicode(base_url + "/".join([MEDIA_URL, 'documentos', 'userreports', elimina_tildes(request.user.username), pdfname + "." + tipo]))
                                }
                    d = run_report_v1(reporte=reporte, tipo=tipo, paRequest=paRequest, request=request)
                    if not d['isSuccess']:
                        raise NameError(d['mensaje'])
                    else:
                        data['archivo']= archivo = d['data']['reportfile']
                        data['url_archivo'] = '{}{}'.format(dominio_sistema, archivo)
                        data['id_objeto'] = cronograma_id
                        data['action_firma']='firmaracta'
                        template = get_template("formfirmaelectronica.html")
                        return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return bad_json(mensaje="Error, al generar el reporte. %s" % ex.__str__())

            elif action == 'reportconstatacion':
                try:
                    titulo = 'Generando reporte de constatación de activos.'
                    noti = Notificacion(cuerpo='Reporte de constatación de activos en progreso',
                                        titulo=titulo, destinatario=personaactivo,
                                        url='',
                                        prioridad=1, app_label='sga-sagest',
                                        fecha_hora_visible=datetime.now() + timedelta(days=1), tipo=2,
                                        en_proceso=True)
                    noti.save(request)
                    reporte_activos_tecnologicos_constatados_openxl_background(request=request, data=data, notif=noti.pk).start()

                    return HttpResponseRedirect(f'{request.path}?action=reporteconstatacion&id={request.GET["id_obj"]}')
                except Exception as ex:
                    messages.error(request, f'Error: {ex}')
                    return HttpResponseRedirect(f'{request.path}?action=reporteconstatacion&id={request.GET["id_obj"]}')

            return HttpResponseRedirect(request.path)

        else:
            try:
                data['title'] = u'Periodos de inventario'
                filtro, search = Q(status=True), request.GET.get('s','')
                if search:
                    data['s'] = search = request.GET['s'].strip()
                    ss = search.split(' ')
                    if len(ss) == 1:
                        filtro = filtro & ((Q(nombre__icontains=search) | Q(detalle__icontains=search)))
                    else:
                        filtro = filtro & ((Q(nombre__icontains=ss[0]) & Q(nombre__icontains=ss[1])) |
                                            (Q(detalle__icontains=ss[0]) & Q(detalle__icontains=ss[1])))
                periodos = PeriodoInventarioAT.objects.filter(filtro).distinct().order_by('-id')
                paging = MiPaginador(periodos, 25)
                p = 1
                try:
                    paginasesion = 1
                    if 'paginador' in request.session:
                        paginasesion = int(request.session['paginador'])
                    if 'page' in request.GET:
                        p = int(request.GET['page'])
                    else:
                        p = paginasesion
                    try:
                        page = paging.page(p)
                    except:
                        p = 1
                    page = paging.page(p)
                except:
                    page = paging.page(p)
                request.session['paginador'] = p
                data['paging'] = paging
                data['rangospaging'] = paging.rangos_paginado(p)
                data['page'] = page
                data['hoy']=datetime.now()
                data['listadocatalogo'] = page.object_list
                data['totales'] = len(periodos.values('id'))
                return render(request, "inventario_activofijo/periodoinventario.html", data)
            except Exception as ex:
                pass