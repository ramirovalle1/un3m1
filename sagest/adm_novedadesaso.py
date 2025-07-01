# -*- coding: UTF-8 -*-
import random
import sys

import xlsxwriter
import io
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Q
from django.db.models.functions import ExtractYear
import xlwt
from django.forms import model_to_dict
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.template.loader import get_template
from openpyxl import load_workbook
from xlwt import *
from django.shortcuts import render
from decorators import secure_module
from sagest.forms import DepartamentoForm, IntegranteDepartamentoForm, ResponsableDepartamentoForm, \
    SeccionDepartamentoForm, ProductoServicioSeccionForm, DepartamentoProductosTHForm, SubirSubNovedadesForm, \
    NotificarForm, ResponsableNovedadForm
from sagest.models import Departamento, SeccionDepartamento, ProductoServicioSeccion, ProductoServicioTh, \
    DetalleSubnovedadPeriodoRol, DetallePeriodoRol, ResponsableNovedad, PeriodoRol
from settings import EMAIL_DOMAIN
from sga.commonviews import adduserdata
from sga.funciones import MiPaginador, log, generar_nombre, remover_caracteres_especiales_unicode
from sga.models import Administrativo, ModuloGrupo, Persona, MONTH_CHOICES, CUENTAS_CORREOS
from sga.tasks import send_html_mail


@login_required(redirect_field_name='ret', login_url='/loginsagest')
#@secure_module
@transaction.atomic()
def view(request):
    data = {}
    adduserdata(request, data)
    usuario = request.user
    persona = request.session['persona']

    if request.method == 'POST':
        action = request.POST['action']

        if action == 'importar':
            try:
                f = SubirSubNovedadesForm(request.POST)
                if f.is_valid():
                    nfile = request.FILES['archivo']
                    rubrorol = f.cleaned_data['rubrorol']
                    archivo = io.BytesIO(nfile.read())
                    workbook = load_workbook(filename=archivo, read_only=False)
                    libro = workbook[workbook.sheetnames[0]]
                    for rowx in range(2, libro.max_row + 1):
                        cedula = (libro.cell(row=rowx, column=1).value).strip()
                        observacion = libro.cell(row=rowx, column=2).value.strip()
                        valor = float(libro.cell(row=rowx, column=3).value)
                        anio = libro.cell(row=rowx, column=4).value
                        mes = libro.cell(row=rowx, column=5).value
                        socio = Persona.objects.filter((Q(cedula__icontains=str(cedula)) |
                                                        Q(pasaporte__icontains=cedula)), status=True).first()
                        if socio:
                            detalleperiodorol = DetallePeriodoRol.objects.filter(persona=socio,
                                                                                 rubro=rubrorol, periodo__anio=anio,
                                                                                 periodo__mes=mes, status=True,
                                                                                 periodo__status=True
                                                                                 ).distinct().first()
                            if detalleperiodorol:
                                detallesubnovedad = detalleperiodorol.detallesubnovedadperiodorol_set.filter(valor=valor,
                                                                                                             descripcion=observacion,
                                                                                                             status=True).first()
                                if detallesubnovedad is None:
                                    detallesubnovedad = DetalleSubnovedadPeriodoRol(
                                        detalleperiodorol=detalleperiodorol,
                                        descripcion=observacion,
                                        valor=valor)
                                    detallesubnovedad.save(request)
                    log(u'Importó archivos con subnovedades de rol: %s' % persona, request, "add")
                    return JsonResponse({"result": False}, safe=False)
                else:
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "message": "Error en el formulario"})
            except Exception as ex:
                linea_ = sys.exc_info()[-1].tb_lineno
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        if action == 'addresponsable':
            try:
                f = ResponsableNovedadForm(request.POST)
                if f.is_valid():
                    if ResponsableNovedad.objects.filter(status= True, rubro= f.cleaned_data['rubro'], persona= f.cleaned_data['persona']):
                        return JsonResponse({"result": True, "mensaje": "Este registro ya existe"}, safe=False)

                    responsable = ResponsableNovedad(
                        persona= f.cleaned_data['persona'],
                        rubro=f.cleaned_data['rubro']
                    )
                    responsable.save(request)
                    if 'logo' in request.FILES:
                        newfile = request.FILES['logo']
                        newfile._name = 'logo_{}'.format(generar_nombre(remover_caracteres_especiales_unicode(responsable.rubro.abreviatura), newfile._name))
                        newfile.logo = newfile
                        responsable.logo = newfile
                        responsable.save(request)
                    log(u'Adicionó responsable: %s' % persona, request, "add")
                    return JsonResponse({"result": False}, safe=False)
                else:
                    return JsonResponse({"result": True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "mensaje": "Error en el formulario"})
            except Exception as ex:
                msg_ex = 'Error on line {} - {}'.format(sys.exc_info()[-1].tb_lineno, str(ex))
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Inténtelo más tarde. %s" % msg_ex}, safe=False)

        if action == 'editresponsable':
            try:
                f = ResponsableNovedadForm(request.POST)
                if f.is_valid():
                    responsable = ResponsableNovedad.objects.get(pk=request.POST['id'])
                    if ResponsableNovedad.objects.filter(status=True, rubro=f.cleaned_data['rubro'], persona=f.cleaned_data['persona']):
                        return JsonResponse({"result": True, "mensaje": "Este registro ya existe"}, safe=False)

                    responsable.persona = f.cleaned_data['persona']
                    responsable.rubro = f.cleaned_data['rubro']
                    if 'logo' in request.FILES:
                        newfile = request.FILES['logo']
                        newfile._name = 'logo_{}'.format(generar_nombre(remover_caracteres_especiales_unicode(responsable.rubro.abreviatura), newfile._name))
                        responsable.logo = newfile
                    responsable.save(request)
                    log(u'Editó responsable de rol: %s' % persona, request, "add")
                    return JsonResponse({"result": False}, safe=False)
                else:
                    return JsonResponse({"result": True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "mensaje": "Error en el formulario"})
            except Exception as ex:
                msg_ex = 'Error on line {} - {}'.format(sys.exc_info()[-1].tb_lineno, str(ex))
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Inténtelo más tarde. %s" % msg_ex}, safe=False)

        if action == 'delresponsable':
            try:
                with transaction.atomic():
                    responsable = ResponsableNovedad.objects.get(pk=request.POST['id'])
                    responsable.status = False
                    responsable.save(request)
                    log(u'Eliminó responsable %s - la persona: %s' % (responsable, persona), request, "del")
                    return JsonResponse({"result":False}, safe=False)
            except Exception as ex:
                msg_ex = 'Error on line {} - {}'.format(sys.exc_info()[-1].tb_lineno, str(ex))
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": u"Error. Detalle: %s" % msg_ex}, safe=False)

        if action == 'notificar':
            try:
                f = NotificarForm(request.POST)
                if f.is_valid():
                    anio = f.cleaned_data['anio']
                    mes = f.cleaned_data['mes']
                    filtro = Q(status=True, periodo__mes=mes, periodo__anio=anio)
                    detper = DetallePeriodoRol.objects.filter(filtro).distinct()
                    idpersonas = detper.values_list('persona_id', flat=True).distinct()
                    personas = Persona.objects.filter(pk__in=idpersonas)
                    responsable = ResponsableNovedad.objects.get(status=True, persona=persona)
                    for p in personas:
                        subnovedades = DetalleSubnovedadPeriodoRol.objects.filter(detalleperiodorol__in= detper.filter(persona=p).values_list('id',flat=True))
                        if subnovedades.__len__() > 0:
                            #envio de correo
                            saludo = 'Estimada' if p.sexo.id == 1 else 'Estimado'
                            asunto = u"Detalle de descuentos"
                            template = 'adm_novedadesaso/correo_notificacion.html'
                            datos_email = {'sistema': 'AEA',
                                           'subnovedades': subnovedades,
                                           'mes':mes,
                                           'anio':anio,
                                           'persona': p,
                                           'imagen': responsable.logo,
                                           'saludo': saludo}
                            lista_email = [p.emailinst]
                            documentolista = []
                            send_html_mail(asunto, template, datos_email, lista_email, [], documentolista, cuenta=CUENTAS_CORREOS[1][1])
                    return JsonResponse({"result": False}, safe=False)
                else:
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                 "message": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result':True, 'mensaje': f'Error de conexión. {ex.__str__()}'})

        # if action == 'notificar':
        #     try:
        #         f = NotificarForm(request.POST)
        #         if f.is_valid():
        #             anio = f.cleaned_data['anio']
        #             mes = f.cleaned_data['mes']
        #             periodos = PeriodoRol.objects.values_list('id', flat=True).filter(anio=anio, mes=mes, status=True)
        #             filtro = Q (periodo_id__in=periodos, status=True)
        #             detper = DetallePeriodoRol.objects.filter(filtro)
        #             idpersonas = detper.values_list('persona_id', flat=True).distinct()
        #             personas = Persona.objects.filter(pk__in=idpersonas)
        #             for p in personas:
        #                 subnovedades = DetalleSubnovedadPeriodoRol.objects.filter(detalleperiodorol__in= detper.filter(persona=p).values_list('id',flat=True))
        #                 #envio de correo
        #                 saludo = 'Estimada' if p.sexo.id == 1 else 'Estimado'
        #                 asunto = u"NOTIFICACIÓN"
        #                 template = 'adm_novedadesaso/correo_notificacion.html'
        #                 datos_email = {'sistema': 'SGA UNEMI',
        #                                'titulo': titulo,
        #                                'subnovedades': subnovedades,
        #                                'persona': p,
        #                                'saludo': saludo}
        #                 lista_email = [p.emailinst]
        #                 documentolista = []
        #                 send_html_mail(asunto, template, datos_email, lista_email, [], documentolista, cuenta=CUENTAS_CORREOS[4][1])
        #             return JsonResponse({"result": False}, safe=False)
        #         else:
        #             return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
        #                          "message": "Error en el formulario"})
        #     except Exception as ex:
        #         transaction.set_rollback(True)
        #         return JsonResponse({'result':True, 'mensaje': f'Error de conexión. {ex.__str__()}'})

        return JsonResponse({"result": "bad", "mensaje": u"Solicitud Incorrecta."})

    else:
        if 'action' in request.GET:
            data['action'] = action = request.GET['action']

            if action == 'importar':
                try:
                    form = SubirSubNovedadesForm()
                    form.cargar_rubrorol(persona.id)
                    data['form'] = form
                    template = get_template('adm_novedadesaso/modal/formimportar.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'mensaje': ex.__str__()})

            if action == 'addresponsable':
                try:
                    form = ResponsableNovedadForm()
                    form.fields['persona'].queryset=Persona.objects.none()
                    data['form'] = form
                    template = get_template('adm_novedadesaso/modal/formimportar.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'mensaje': ex.__str__()})

            if action == 'editresponsable':
                try:
                    data['filtro'] = responsable = ResponsableNovedad.objects.get(pk=request.GET['id'])
                    form = ResponsableNovedadForm(initial=model_to_dict(responsable))
                    form.fields['persona'].queryset=Persona.objects.filter(id=responsable.persona.id)
                    data['form'] = form
                    template = get_template('adm_novedadesaso/modal/formimportar.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'mensaje': ex.__str__()})

            if action == 'buscarpersona':
                try:
                    q = request.GET['q'].upper().strip()
                    s = q.split(" ")
                    if len(s) == 1:
                        per = Persona.objects.filter((Q(distributivopersona__isnull=False) | Q(profesor__isnull=False)),
                                                     (Q(nombres__icontains=q) | Q(apellido1__icontains=q) | Q(
                                                         apellido2__icontains=q) | Q(cedula__contains=q)),
                                                     Q(status=True)).distinct()[:15]
                    elif len(s) == 2:
                        per = Persona.objects.filter((Q(distributivopersona__isnull=False) | Q(profesor__isnull=False)),
                                                     (Q(apellido1__contains=s[0]) & Q(apellido2__contains=s[1])) | (
                                                             Q(nombres__icontains=s[0]) & Q(
                                                         nombres__icontains=s[1])) | (
                                                             Q(nombres__icontains=s[0]) & Q(
                                                         apellido1__contains=s[1]))).filter(status=True).distinct()[
                              :15]
                    else:
                        per = Persona.objects.filter((Q(distributivopersona__isnull=False) | Q(profesor__isnull=False)),
                                                     (Q(nombres__contains=s[0]) & Q(apellido1__contains=s[1]) & Q(
                                                         apellido2__contains=s[2])) | (Q(nombres__contains=s[0]) & Q(
                                                         nombres__contains=s[1]) & Q(apellido1__contains=s[2]))).filter(
                            status=True).distinct()[:15]

                    data = {"result": "ok",
                            "results": [{"id": x.id, "name": str(x.nombre_completo()), "ci":str(x.documento()), "foto": x.get_foto()}
                                        for x in per]}
                    return JsonResponse(data)
                except Exception as ex:
                    pass

            if action == 'responsable':
                try:
                    data['title'] = u'Productos y servicios'
                    search, url_vars = request.GET.get('s', ''), f'&action={action}'
                    if search:
                        ss = search.split(' ')
                        if len(ss) == 1:
                            responsable = ResponsableNovedad.objects.filter(Q(persona__apellido1__icontains=search) |
                                                                            Q(persona__cedula=search) |
                                                                            Q(persona__pasaporte=search) |
                                                                            Q(persona__nombres__icontains=search), status= True).distinct()
                        else:
                            responsable = ResponsableNovedad.objects.filter(Q(persona__apellido1__icontains=ss[0]) |
                                                                            Q(persona__apellido2__icontains=ss[1]),
                                                                             status=True).distinct()
                    else:
                        responsable = ResponsableNovedad.objects.filter(status=True)
                    paging = MiPaginador(responsable, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['s'] = search if search else ""
                    data['responsables'] = page.object_list
                    data['url_vars'] = url_vars
                    return render(request, "adm_novedadesaso/viewresponsable.html", data)
                except Exception as ex:
                    pass

            if action == 'notificar':
                try:
                    data['form'] = form = NotificarForm()
                    template = get_template('adm_novedadesaso/modal/notificar.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'mensaje': ex.__str__()})

            return HttpResponseRedirect(request.path)

        else:
            data['title'] = u'Novedades para descuento'
            search, messelect, url_vars = request.GET.get('s', ''), \
                                                         request.GET.get('m', 0), ''
            if messelect:
                data['messelect'] = messelect = int(request.GET['m'])
                url_vars += "&messelect={}".format(messelect)

            rubros = ResponsableNovedad.objects.filter(status=True,persona=persona).values_list('rubro_id', flat=True)
            subnovedades = DetalleSubnovedadPeriodoRol.objects.filter(status=True,detalleperiodorol__rubro_id__in=rubros)

            data['meses'] = MONTH_CHOICES
            if 's' in request.GET:
                search = request.GET['s']
                subnovedades = subnovedades.filter(Q(detalleperiodorol__persona__nombres__unaccent__icontains=search)|Q(detalleperiodorol__persona__apellido1__unaccent__icontains=search))
            if 'm' in request.GET:
                mes = int(request.GET['m'])
                if mes>0:
                    subnovedades = subnovedades.filter(detalleperiodorol__periodo__mes=mes)

            paging = MiPaginador(subnovedades, 20)
            p = 1
            try:
                paginasesion = 1
                if 'paginador' in request.session:
                    paginasesion = int(request.session['paginador'])
                if 'page' in request.GET:
                    p = int(request.GET['page'])
                else:
                    p = paginasesion
                try:
                    page = paging.page(p)
                except:
                    p = 1
                page = paging.page(p)
            except:
                page = paging.page(p)
            request.session['paginador'] = p
            data['paging'] = paging
            data['rangospaging'] = paging.rangos_paginado(p)
            data['page'] = page
            data['search'] = search if search else ""
            data['subnovedades'] = page.object_list
            data['email_domain'] = EMAIL_DOMAIN
            return render(request, 'adm_novedadesaso/view.html', data)