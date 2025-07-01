# -*- coding: UTF-8 -*-
import io
import json
import random
import sys
from datetime import datetime
from openpyxl import workbook as openxl
from openpyxl.styles import Font as openxlFont
from openpyxl.styles.alignment import Alignment as alin
import xlsxwriter
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Q
import xlwt
from django.forms import model_to_dict
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.template import Context
from django.template.loader import get_template
from django.views.decorators.csrf import csrf_exempt
from xlwt import *
from django.shortcuts import render

from bd.models import IncidenciaSCRUM
from decorators import secure_module, last_access
from sagest.models import Departamento, SeccionDepartamento
from settings import EMAIL_DOMAIN

from sga.commonviews import adduserdata

from sga.funciones import MiPaginador, log, generar_nombre, notificacion
from sga.models import Persona
from sga.templatetags.sga_extras import encrypt
from utils.filtros_genericos import filtro_persona_select
from .forms import PlanificacionAutomatizaForm, RequerimientoPlanificacionAutomatizaForm, \
    RequerimientoPlanificacionGestorForm
from .models import *

# @csrf_exempt
@login_required(redirect_field_name='ret', login_url='/loginsagest')
@secure_module
@last_access
@transaction.atomic()
def view(request):
    data = {}
    adduserdata(request, data)
    usuario = request.user
    persona = request.session['persona']
    hoy = datetime.now()
    if request.method == 'POST':
        action = request.POST['action']

        #PLANIFICACION
        if action == 'addplanificacion':
            try:
                with transaction.atomic():
                    form = PlanificacionAutomatizaForm(request.POST)
                    if form.is_valid():
                        planificacion = PlanificacionAutomatiza(fechainicio=form.cleaned_data['fechainicio'],
                                                                fechafin=form.cleaned_data['fechafin'],
                                                                nombre=form.cleaned_data['nombre'],
                                                                detalle=form.cleaned_data['detalle'],
                                                                mostrar=form.cleaned_data['mostrar'],
                                                                departamento=form.cleaned_data['departamento'],
                                                                )
                        planificacion.save(request)
                        log(u'Adicionó planificación para automatización: %s' % planificacion, request, "add")
                        # return JsonResponse({"result": False,'to':'/nuevaurl'}, safe=False) SI DESEAS REDIRECCIONAR ADICIONARLE TO A LA RESPUESTA
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Complete los datos requeridos."}, safe=False)
            except Exception as ex:
                pass

        if action == 'editplanificacion':
            try:
                with transaction.atomic():
                    form = PlanificacionAutomatizaForm(request.POST)
                    id=int(encrypt(request.POST['id']))
                    if form.is_valid():
                        planificacion = PlanificacionAutomatiza.objects.get(id=id)
                        planificacion.departamento=form.cleaned_data['departamento']
                        planificacion.fechainicio=form.cleaned_data['fechainicio']
                        planificacion.fechafin=form.cleaned_data['fechafin']
                        planificacion.nombre=form.cleaned_data['nombre']
                        planificacion.detalle=form.cleaned_data['detalle']
                        planificacion.mostrar=form.cleaned_data['mostrar']
                        planificacion.save(request)
                        log(u'Adicionó planificación para automatización: %s' % planificacion, request, "add")
                        # return JsonResponse({"result": False,'to':'/nuevaurl'}, safe=False) SI DESEAS REDIRECCIONAR ADICIONARLE TO A LA RESPUESTA
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Complete los datos requeridos."}, safe=False)
            except Exception as ex:
                pass

        if action == 'delplanificacion':
            with transaction.atomic():
                try:
                    instancia = PlanificacionAutomatiza.objects.get(pk=int(encrypt(request.POST['id'])))
                    instancia.status = False
                    instancia.save(request)
                    log(u'Elimino planificación: %s' % instancia, request, "del")
                    res_json = {"error": False}
                except Exception as ex:
                    res_json = {'error': True, "message": "Error: {}".format(ex)}
                return JsonResponse(res_json, safe=False)

        if action == 'mostrarrequerimiento':
            with transaction.atomic():
                try:
                    mostrar = eval(request.POST['val'].capitalize())
                    registro = PlanificacionAutomatiza.objects.get(pk=int(request.POST['id']))
                    registro.mostrar = mostrar
                    registro.save(request)
                    log(u'Mostrar requerimiento : %s (%s)' % (registro, registro.mostrar), request,"edit")
                    return JsonResponse({"result": True, 'mensaje': 'Cambios guardados'})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": False})

        # REQUERIMIETNOS DE PLANIFICACION
        if action == 'addrequerimiento':
            with transaction.atomic():
                try:
                    idp = int(encrypt(request.POST['idp']))
                    form = RequerimientoPlanificacionGestorForm(request.POST)
                    if not form.is_valid():
                        transaction.set_rollback(True)
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                             "mensaje": "Error en el formulario"})

                    requerimiento = RequerimientoPlanificacionAutomatiza(periodo_id=idp,
                                                                         gestion=form.cleaned_data['gestion'],
                                                                         prioridad=form.cleaned_data['prioridad'],
                                                                         # orden=form.cleaned_data['orden'],
                                                                         responsable=form.cleaned_data['responsable'],
                                                                         detalle=form.cleaned_data['detalle'],
                                                                         procedimiento=form.cleaned_data['procedimiento'],
                                                                         )
                    requerimiento.save(request)
                    documentos = request.FILES.getlist('adjuntos')
                    lista_items1 = json.loads(request.POST['lista_items1'])
                    for d in documentos:
                        items = [item for item in lista_items1 if item['archivo'] == d._name]
                        if len(items) > 1 and all(item['size'] == items[0]['size'] for item in items):
                            raise NameError(f'Error, archivos duplicados {d._name}, remplace uno de los archivos duplicados.')
                        if d.size > 4194304:
                            raise NameError(f"Error, archivo es mayor a 4 Mb. {items[0]['descripcion']}")
                        d._name = generar_nombre(f"adjunto_{requerimiento.id}_", d._name)
                        doc = DocumentoAdjuntoRequerimiento(requerimiento=requerimiento, leyenda=items[0]['descripcion'], archivo=d)
                        doc.save(request)
                    log(u'Adicionó requerimiento: %s' % requerimiento, request, "add")

                    actividad = IncidenciaSCRUM(requerimiento=requerimiento,
                                                titulo=form.cleaned_data['procedimiento'],
                                                descripcion=form.cleaned_data['detalle'],
                                                prioridad=requerimiento.prioridad)
                    actividad.save(request)
                    for lider in actividad.lideres_departamento():
                        titulo = 'Nueva incidencia recibida'
                        cuerpo = f'Estimado líder de equipo, se ingresó el requerimiento {actividad}, en caso de no pertenecer a su gestión hacer caso omiso '
                        notificacion(titulo, cuerpo, lider,
                                     None, '/adm_scrum_actividades?action=requerimientos', lider.pk, 2, 'sga-sagest',
                                     IncidenciaSCRUM, request)
                    return JsonResponse({"result": False}, safe=False)
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": True, "mensaje": f"Error: {ex}"}, safe=False)

        if action == 'editrequerimiento':
            with transaction.atomic():
                try:
                    form = RequerimientoPlanificacionGestorForm(request.POST)
                    id = int(encrypt(request.POST['id']))
                    idp = int(encrypt(request.POST['idp']))
                    if form.is_valid():
                        requerimiento = RequerimientoPlanificacionAutomatiza.objects.get(id=id)
                        requerimiento.gestion = form.cleaned_data['gestion']
                        requerimiento.prioridad = form.cleaned_data['prioridad']
                        # requerimiento.orden = form.cleaned_data['orden']
                        requerimiento.responsable = form.cleaned_data['responsable']
                        requerimiento.detalle = form.cleaned_data['detalle']
                        requerimiento.procedimiento = form.cleaned_data['procedimiento']
                        requerimiento.save(request)
                        actividad = requerimiento.incidenciascrum_set.filter(status=True).first()
                        if actividad:
                            actividad.titulo = form.cleaned_data['procedimiento']
                            actividad.descripcion = form.cleaned_data['detalle']
                            actividad.prioridad = requerimiento.prioridad
                            actividad.save(request)
                        documentos = request.FILES.getlist('adjuntos')
                        lista_items1 = json.loads(request.POST['lista_items1'])
                        ids_excl = [int(item['id_adjunto']) for item in lista_items1 if item['id_adjunto']]
                        for d in documentos:
                            items = [item for item in lista_items1 if item['archivo'] == d._name]
                            if len(items) > 1 and all(item['size'] == items[0]['size'] for item in items):
                                raise NameError(f'Error, archivos duplicados {d._name}, remplace uno de los archivos duplicados.')
                            if d.size > 4194304:
                                raise NameError(f"Error, archivo es mayor a 4 Mb. {items[0]['descripcion']}")
                            if not items[0]['id_adjunto']:
                                d._name = generar_nombre(f"adjunto_{requerimiento.id}_", d._name)
                                doc = DocumentoAdjuntoRequerimiento(requerimiento=requerimiento, leyenda=items[0]['descripcion'], archivo=d)
                                doc.save(request)
                            else:
                                doc = DocumentoAdjuntoRequerimiento.objects.get(id=items[0]['id_adjunto'])
                                doc.leyenda = items[0]['descripcion']
                                doc.archivo = d
                                doc.save(request)
                            ids_excl.append(doc.id)
                        for items in lista_items1:
                            if items['id_adjunto']:
                                doc = DocumentoAdjuntoRequerimiento.objects.get(id=int(items['id_adjunto']))
                                doc.leyenda = items['descripcion']
                                doc.save(request, update_fields=['leyenda'])
                        requerimiento.documentos().exclude(id__in=ids_excl).update(status=False)
                        log(u'Edito requerimiento: %s' % requerimiento, request, "edit")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                             "mensaje": "Error en el formulario"})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        if action == 'delrequerimiento':
            with transaction.atomic():
                try:
                    instancia = RequerimientoPlanificacionAutomatiza.objects.get(pk=int(encrypt(request.POST['id'])))
                    instancia.status = False
                    instancia.save(request)
                    log(u'Elimino requerimiento: %s' % instancia, request, "del")
                    res_json = {"error": False}
                except Exception as ex:
                    res_json = {'error': True, "message": "Error: {}".format(ex)}
                return JsonResponse(res_json, safe=False)

        return JsonResponse({"result": "bad", "mensaje": u"Solicitud Incorrecta."})
    else:
        if 'action' in request.GET:
            data['action'] = action = request.GET['action']
            midepartamento_ = persona.mi_departamento()
            if action == 'requerimientos':
                try:
                    data['idp'] = idp = int(encrypt(request.GET['idp']))
                    data['plan'] = plan = PlanificacionAutomatiza.objects.get(id=idp)
                    data['title'] = 'Requerimiento'
                    url_vars, filtro, search, prioridad = f'&action={action}&idp={request.GET["idp"]}', \
                                                          Q(status=True, periodo_id=idp), \
                                                          request.GET.get('s', ''), request.GET.get('prioridad', '')

                    if search:
                        data['s'] = search
                        filtro = filtro & (Q(procedimiento__icontains=search))
                        url_vars += f'&s={search}'
                    if prioridad:
                        data['prioridad'] = pr = int(prioridad)
                        filtro = filtro & (Q(prioridad=int(prioridad)))
                        url_vars += f'&prioridad={pr}'
                    planes = RequerimientoPlanificacionAutomatiza.objects.filter(filtro).order_by('-id')

                    paging = MiPaginador(planes, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['prioridades'] = PRIORIDAD
                    data['search'] = search if search else ""
                    data["url_vars"] = url_vars
                    data['listado'] = page.object_list
                    data['total']=len(planes.values_list('id'))
                    data['email_domain'] = EMAIL_DOMAIN
                    if 'exportar_excel' in request.GET:
                        wb = openxl.Workbook()
                        wb["Sheet"].title = "Reporte_requerimientos"
                        ws = wb.active
                        style_title = openxlFont(name='Arial', size=16, bold=True)
                        style_cab = openxlFont(name='Arial', size=10, bold=True)
                        alinear = alin(horizontal="center", vertical="center")
                        response = HttpResponse(content_type="application/ms-excel")
                        response[
                            'Content-Disposition'] = 'attachment; filename=Reporte de requerimientos' + '-' + random.randint(
                            1, 10000).__str__() + '.xlsx'
                        ws.column_dimensions['B'].width = 25
                        ws.column_dimensions['C'].width = 25
                        ws.column_dimensions['D'].width = 15
                        ws.column_dimensions['E'].width = 25
                        ws.column_dimensions['F'].width = 25
                        ws.column_dimensions['G'].width = 25
                        ws.column_dimensions['I'].width = 25
                        ws.merge_cells('A1:G1')
                        ws['A1'] = 'PLANIFICACIÓN DE REQUERIMIENTOS'
                        celda1 = ws['A1']
                        celda1.font = style_title
                        celda1.alignment = alinear

                        columns = [u"N°", u"PERIODO", u"GESTIÓN", u"PRIORIDAD", u"RESPONSABLE",
                                   u"DETALLE", u"PROCEDIMIENTO"
                                   ]
                        row_num = 3
                        for col_num in range(0, len(columns)):
                            celda = ws.cell(row=row_num, column=(col_num + 1), value=columns[col_num])
                            celda.font = style_cab
                        row_num = 4
                        for list in planes:
                            ws.cell(row=row_num, column=1, value=list.orden)
                            ws.cell(row=row_num, column=2, value=str(list.periodo.nombre))
                            ws.cell(row=row_num, column=3, value=str(list.gestion))
                            ws.cell(row=row_num, column=4, value=str(list.get_prioridad_display()))
                            ws.cell(row=row_num, column=5, value=str(list.responsable.nombre_completo_minus()))
                            ws.cell(row=row_num, column=6, value=str(list.detalle))
                            ws.cell(row=row_num, column=7, value=str(list.procedimiento))
                            row_num += 1
                        wb.save(response)
                        return response
                    return render(request, 'adm_gestionarequerimiento/requerimientos.html', data)
                except Exception as ex:
                    pass

            elif action == 'addplanificacion':
                try:
                    data['id'] = request.GET['id']
                    data['form2'] = PlanificacionAutomatizaForm()
                    template = get_template("adm_gestionarequerimiento/modal/formplanificacion.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'editplanificacion':
                try:
                    data['id'] = request.GET['id']
                    plan=PlanificacionAutomatiza.objects.get(id=int(encrypt(request.GET['id'])))
                    data['form2'] = PlanificacionAutomatizaForm(initial=model_to_dict(plan))
                    template = get_template("adm_gestionarequerimiento/modal/formplanificacion.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'addrequerimiento':
                try:
                    data['idp'] = request.GET['id']
                    req = RequerimientoPlanificacionAutomatiza.objects.filter(status=True, periodo_id=int(encrypt(request.GET['id']))).order_by('orden').last()
                    cantidad = req.orden if req else 0
                    cantidad += 1
                    form = RequerimientoPlanificacionGestorForm()
                    form.fields['responsable'].queryset = Persona.objects.none()
                    form.fields['gestion'].queryset = SeccionDepartamento.objects.none()
                    data['form'] = form
                    data['seccionado'] = True
                    data['cantidad'] = cantidad
                    template = get_template("adm_gestionarequerimiento/modal/formrequerimiento.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'editrequerimiento':
                try:
                    data['filtro'] = requerimiento = RequerimientoPlanificacionAutomatiza.objects.get(
                        pk=int(encrypt(request.GET['id'])))
                    data['idp'] = encrypt(requerimiento.periodo.id)
                    form = RequerimientoPlanificacionGestorForm(initial=model_to_dict(requerimiento))
                    form.fields['responsable'].queryset = Persona.objects.filter(id=requerimiento.responsable.id)
                    form.fields['gestion'].queryset = SeccionDepartamento.objects.filter(departamento_id=requerimiento.gestion.departamento.id)
                    data['form'] = form
                    data['seccionado'] = True
                    template = get_template("adm_gestionarequerimiento/modal/formrequerimiento.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'listgestiones':
                try:
                    lista = []
                    id = int(request.GET['id'])
                    gestiones = SeccionDepartamento.objects.filter(status=True, departamento=id).distinct()
                    for s in gestiones:
                        text = str(s)
                        lista.append({'value': s.id, 'text': text})
                    return JsonResponse({'result': True, 'data': lista})
                except Exception as e:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            elif action == 'buscarpersonas':
                try:
                    resp = filtro_persona_select(request)
                    return HttpResponse(json.dumps({'status': True, 'results': resp}))
                except Exception as ex:
                    pass

            return HttpResponseRedirect(request.path)
        else:
            data['title'] = u'Planificaciones'

            url_vars = ''
            filtro = Q(status=True)
            search = None
            ids = None
            tipo = None

            if 's' in request.GET:
                if request.GET['s'] != '':
                    search = request.GET['s']

            if search:
                filtro = filtro & (Q(nombre__icontains=search))
                url_vars += '&s=' + search


            planificacion = PlanificacionAutomatiza.objects.filter(filtro).order_by('-id')

            paging = MiPaginador(planificacion, 20)
            p = 1
            try:
                paginasesion = 1
                if 'paginador' in request.session:
                    paginasesion = int(request.session['paginador'])
                if 'page' in request.GET:
                    p = int(request.GET['page'])
                else:
                    p = paginasesion
                try:
                    page = paging.page(p)
                except:
                    p = 1
                page = paging.page(p)
            except:
                page = paging.page(p)
            request.session['paginador'] = p
            data['paging'] = paging
            data['rangospaging'] = paging.rangos_paginado(p)
            data['page'] = page
            data['search'] = search if search else ""
            data["url_vars"] = url_vars
            data['ids'] = ids if ids else ""
            data['planificaciones'] = page.object_list
            data['email_domain'] = EMAIL_DOMAIN
            return render(request, 'adm_gestionarequerimiento/view.html', data)