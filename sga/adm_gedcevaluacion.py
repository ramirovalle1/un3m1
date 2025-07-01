# -*- coding: UTF-8 -*-
import json
import random
import sys

import statistics as stats
import openpyxl
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Q
import xlwt
from django.forms import model_to_dict
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.template import Context
from django.template.loader import get_template
from xlwt import *
from django.shortcuts import render, redirect
from decorators import secure_module, last_access
from settings import EMAIL_DOMAIN
from sga.commonviews import adduserdata
from .forms import GEDCIndicadoresForm, GEDCCabForm, GEDCPreguntasForm, GEDCPreguntasEditarForm, GEDCFactorForm
from sga.funciones import MiPaginador, log, generar_nombre, remover_caracteres_especiales_unicode
from sga.models import Administrativo, Persona, GEDCFactores, InstitucionEducacionSuperior, GEDC_GRUPO, \
    GENEROS_ENCUESTAS
from .models import GEDCCab, GEDCIndicador, GEDCPreguntas, GEDCPersona, GEDCRespuestas, TIPO_CALCULO_GEDC, GEDC_DETALLE_BUSQUEDA
from django.contrib.postgres.aggregates import ArrayAgg


@login_required(redirect_field_name='ret', login_url='/loginsga')
@secure_module
@last_access
@transaction.atomic()
def view(request):
    data = {}
    adduserdata(request, data)
    usuario = request.user
    if request.method == 'POST':
        res_json = []
        action = request.POST['action']

        if action == 'cargarExcelBancoPregunta':
            try:
                with transaction.atomic():
                    excel = request.FILES['excel']
                    wb = openpyxl.load_workbook(excel)
                    worksheet = wb.worksheets[0]
                    count = 0
                    counter = 0
                    linea = 1
                    for row in worksheet.iter_rows():
                        currentValues = [str(cell.value) for cell in row]
                        if linea >= 2:
                            if currentValues[0] == 'None':
                                messages.error(request, 'REVISAR LINEA [{}],  FORMATO FILAS SIN CODIGO DE GRUPO, VERIFIQUE QUE EL ARCHIVO NO TENGA COLUMNAS VACIAS AL FINAL.'.format(linea))
                                transaction.set_rollback(True)
                                return redirect('{}?action=configuraciones'.format(request.path))
                            if currentValues[1] == 'None':
                                messages.error(request, 'REVISAR LINEA [{}],  FORMATO FILAS SIN NOMBRE DE MATERIA, VERIFIQUE QUE EL ARCHIVO NO TENGA COLUMNAS VACIAS AL FINAL.'.format(linea))
                                transaction.set_rollback(True)
                                return redirect('{}?action=configuraciones'.format(request.path))
                            if currentValues[1] == '':
                                messages.error(request, 'REVISAR LINEA [{}],  LA COLUMNA DE DESCRIPCION DEBE ESTAR LLENA.'.format(linea))
                                transaction.set_rollback(True)
                                return redirect('{}?action=configuraciones'.format(request.path))
                            grupo = currentValues[0] if currentValues[0] != 'None' else ''
                            descripcion = currentValues[1] if currentValues[1] != 'None' else ''
                            evalua = currentValues[2] if currentValues[2] != 'None' else ''
                            evalua = True if evalua == 'SI' else False
                            calificacion = currentValues[3] if currentValues[3] != 'None' else ''
                            calificacion = True if calificacion == 'SI' else False
                            observacion = currentValues[4] if currentValues[4] != 'None' else ''
                            observacion = True if observacion == 'SI' else False
                            evidencias = currentValues[5] if currentValues[5] != 'None' else ''
                            evidencias = True if evidencias == 'SI' else False
                            if GEDCIndicador.objects.filter(grupo=grupo, status=True, indicador__icontains=descripcion).exists():
                                count += 1
                            else:
                                car = GEDCIndicador(indicador=descripcion,
                                                    grupo=grupo,
                                                    evalua=evalua,
                                                    calificacion=calificacion,
                                                    observacion=observacion,
                                                    evidencias=evidencias,)
                                car.save(request)
                                log(u'Adiciono Pregunta al Banco de Preguntas GEDC: %s' % car, request, "add")
                                counter += 1
                        linea += 1
                    if count > 0:
                        messages.info(request, 'Se actualizaron {} registros ya existentes.'.format(str(count)))
                    if counter > 0:
                        messages.success(request, 'Se agregaron {} preguntas'.format(counter))
            except Exception as ex:
                print(ex)
                print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
                messages.error(request, ex)
                messages.error(request, 'Acaba de ocurrir un error al intertar cargar {}, Acción fue cancelada.'.format(currentValues))
            return redirect('{}?action=configuraciones'.format(request.path))

        if action == 'addencuesta':
            try:
                f = GEDCCabForm(request.POST)
                if f.is_valid():
                    nombre_url = remover_caracteres_especiales_unicode(request.POST['nombre']).lower().replace('\t','').replace(' ', '_')
                    if GEDCCab.objects.filter(nombreurl=nombre_url).exists():
                        transaction.set_rollback(True)
                        return JsonResponse({"result": "bad", "mensaje": u"Ya existe una encuesta con ese nombre %s" % nombre_url})
                    filtro = GEDCCab(nombre=f.cleaned_data['nombre'],
                                     grupo = f.cleaned_data['grupo'],
                                     detalle=f.cleaned_data['detalle'],
                                     publicar =f.cleaned_data['publicar'])
                    filtro.nombreurl = nombre_url
                    filtro.save(request)
                    messages.success(request,'REGISTRO GUARDADO')
                    log(u'Adiciono nueva Encuesta GEDC: %s' % filtro, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos. %s" % ex})

        if action == 'editencuesta':
            try:
                f = GEDCCabForm(request.POST)
                if f.is_valid():
                    filtro = GEDCCab.objects.get(pk=int(request.POST['id']))
                    filtro.nombre=f.cleaned_data['nombre']
                    filtro.grupo = f.cleaned_data['grupo']
                    filtro.detalle = f.cleaned_data['detalle']
                    filtro.publicar = f.cleaned_data['publicar']
                    filtro.save(request)
                    log(u'Editó Encuesta GEDC: %s' % filtro, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos. %s" % ex})

        if action == 'delencuesta':
            try:
                filtro = GEDCCab.objects.get(pk=request.POST['id'])
                filtro.status = False
                filtro.save(request)
                log(u'Elimino Encuesta GEDC: %s' % filtro, request, "del")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        if action == 'addpregunta':
            try:
                with transaction.atomic():
                    idcab = int(request.POST['id'])
                    cabpregunta = GEDCCab.objects.get(pk=idcab)
                    if GEDCPreguntas.objects.filter(cab=cabpregunta, indicador_id=int(request.POST['indicador'])):
                        res_json.append({'result': True, "message": 'Indicador ya registrado'})
                        return JsonResponse(res_json, safe=False)
                    form = GEDCPreguntasForm(request.POST)
                    if form.is_valid():
                        filtro = GEDCPreguntas(cab=cabpregunta,
                                               indicador=form.cleaned_data['indicador'],
                                               orden=form.cleaned_data['orden'],
                                               obligatorio=form.cleaned_data['obligatorio'])
                        filtro.save(request)
                        log(u'Adiciono indicador a la encuesta gedc: %s' % cabpregunta, request, "add")
                        # return JsonResponse({"result": False,'to':'/nuevaurl'}, safe=False) SI DESEAS REDIRECCIONAR ADICIONARLE TO A LA RESPUESTA
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Complete los datos requeridos."}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        if action == 'editpregunta':
            try:
                with transaction.atomic():
                    idcab = int(request.POST['id'])
                    cabpregunta = GEDCPreguntas.objects.get(pk=idcab)
                    form = GEDCPreguntasEditarForm(request.POST)
                    if form.is_valid():
                        cabpregunta.orden=form.cleaned_data['orden']
                        cabpregunta.obligatorio=form.cleaned_data['obligatorio']
                        cabpregunta.save(request)
                        log(u'Edito indicador de la encuesta gedc: %s' % cabpregunta, request, "add")
                        # return JsonResponse({"result": False,'to':'/nuevaurl'}, safe=False) SI DESEAS REDIRECCIONAR ADICIONARLE TO A LA RESPUESTA
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Complete los datos requeridos."}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        if action == 'delpregunta':
            try:
                filtro = GEDCPreguntas.objects.get(pk=request.POST['id'])
                log(u'Elimino Pregunta GEDC: %s' % filtro, request, "del")
                filtro.delete()
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        if action == 'addindicador':
            try:
                with transaction.atomic():
                    form = GEDCIndicadoresForm(request.POST)
                    if form.is_valid():
                        filtro = GEDCIndicador(indicador=form.cleaned_data['indicador'].upper(),
                                              factores=form.cleaned_data['factores'],
                                              grupo=form.cleaned_data['grupo'],
                                              sentido=form.cleaned_data['sentido'],
                                              evalua=form.cleaned_data['evalua'],
                                              calificacion=form.cleaned_data['calificacion'],
                                              observacion=form.cleaned_data['observacion'],
                                              evidencias=form.cleaned_data['evidencias'])
                        filtro.save(request)
                        log(u'Adiciono indicador al banco de preguntas gedc: %s' % filtro, request, "add")
                        # return JsonResponse({"result": False,'to':'/nuevaurl'}, safe=False) SI DESEAS REDIRECCIONAR ADICIONARLE TO A LA RESPUESTA
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Complete los datos requeridos."}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        if action == 'editindicador':
            try:
                with transaction.atomic():
                    filtro = GEDCIndicador.objects.get(pk=request.POST['id'])
                    f = GEDCIndicadoresForm(request.POST)
                    if f.is_valid():
                        filtro.indicador = f.cleaned_data['indicador'].upper()
                        filtro.sentido = f.cleaned_data['sentido']
                        filtro.factores = f.cleaned_data['factores']
                        filtro.grupo = f.cleaned_data['grupo']
                        filtro.evalua = f.cleaned_data['evalua']
                        filtro.calificacion = f.cleaned_data['calificacion']
                        filtro.observacion = f.cleaned_data['observacion']
                        filtro.evidencias = f.cleaned_data['evidencias']
                        filtro.save(request)
                        log(u'Modificó indicador en el banco de preguntas gedc: %s' % filtro, request, "edit")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Complete los datos requeridos."}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        if action == 'delindicador':
            try:
                filtro = GEDCIndicador.objects.get(pk=request.POST['id'])
                filtro.status = False
                filtro.save(request)
                log(u'Elimino indicador en el banco de preguntas gedc: %s' % filtro, request, "del")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        if action == 'addfactor':
            try:
                with transaction.atomic():
                    form = GEDCFactorForm(request.POST)
                    if form.is_valid():
                        filtro = GEDCFactores(nombres=form.cleaned_data['nombres'].upper())
                        filtro.save(request)
                        log(u'Adiciono factor gedc: %s' % filtro, request, "add")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Complete los datos requeridos."}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        if action == 'editfactor':
            try:
                with transaction.atomic():
                    filtro = GEDCFactores.objects.get(pk=request.POST['id'])
                    f = GEDCFactorForm(request.POST)
                    if f.is_valid():
                        filtro.nombres = f.cleaned_data['nombres'].upper()
                        filtro.save(request)
                        log(u'Modificó factor gedc: %s' % filtro, request, "edit")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Complete los datos requeridos."}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        if action == 'delfactor':
            try:
                filtro = GEDCFactores.objects.get(pk=request.POST['id'])
                filtro.status = False
                filtro.save(request)
                log(u'Elimino factor gedc: %s' % filtro, request, "del")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        if action == 'estadistica':
            try:
                idgrupo = int(request.POST['grupo'])
                if idgrupo > 0:
                    qsgrupo = GEDCCab.objects.filter(status=True, pk=idgrupo)
                else:
                    qsgrupo = GEDCCab.objects.filter(status=True, publicar=True)
                qspreguntas = GEDCPreguntas.objects.filter(status=True, cab__in=qsgrupo.values_list('id', flat=True))
                return JsonResponse({"result": True}, safe=False)
            except Exception as ex:
                return JsonResponse({"result": False, "mensaje": str(ex)}, safe=False)

        if action == 'listResultados':
            try:
                criterio, desde, hasta, listCount = request.POST.get('sSearch', ''), int(request.POST.get('iDisplayStart', 0)), int(request.POST.get('iDisplayLength', 0)), 0
                grupo, pais, universidad, variable, calculo = int(request.POST.get('grupo','')), request.POST.get('pais',''), request.POST.get('universidad',''), request.POST.get('variable',''), request.POST.get('calculo',''),
                porgrupo, porpais, poruniversidad, porgenero = False, False, False, False
                hasta = hasta if desde == 0 else desde+hasta
                if variable:
                    listavariables = variable.split(',')
                    for lv in listavariables:
                        if int(lv) == 1:
                            porgrupo = True
                        if int(lv) == 2:
                            porpais = True
                        if int(lv) == 3:
                            poruniversidad = True
                        if int(lv) == 4:
                            porgenero = True
                aaData = []
                if not porgrupo and not porpais and not poruniversidad and not porgenero:
                    filtro, listarespuestas = Q(status=True), []
                    if grupo:
                        if grupo != 0:
                            filtro = filtro & (Q(cab__grupo=grupo))
                    if pais:
                        if pais != '0':
                            listapais = list(GEDCRespuestas.objects.select_related('cab').filter(Q(cab__pais__nombre__icontains=pais)).values_list('pregunta__id',flat=True))
                            listarespuestas += listapais
                    if universidad:
                        if universidad != '0' or universidad != 'null':
                            listauniversidad = list(GEDCRespuestas.objects.select_related('cab').filter(Q(cab__universidad__nombre__icontains=universidad) | Q(cab__otrauniversidad__icontains=universidad)).values_list('pregunta__id',flat=True))
                            listarespuestas += listauniversidad
                    if listarespuestas:
                        filtro = filtro & Q(id__in=listarespuestas)
                    if criterio:
                        search = criterio.strip()
                        filtro = filtro &  (Q(indicador__indicador__icontains=search) | Q(indicador__factores__nombres__icontains=search))
                    listado = GEDCPreguntas.objects.filter(cab__publicar=True).filter(filtro).order_by('pk')
                    listCount = listado.count()
                    listado = listado[desde:hasta]
                    for row in listado:
                        aaData.append([row.cab.get_grupo_display(), '', '', '', row.get_factor(), row.indicador.indicador, row.total_respuestas(), round(row.get_media(),2), round(row.get_desviacionestandar(),2), porgrupo, porpais, poruniversidad, porgenero])
                else:
                    filtro, listDistinct, listOrderBy, valuesList = Q(status=True), [], [], []
                    if porgrupo:
                        listDistinct.append('cab__cab__grupo')
                        listOrderBy.append('cab__cab__grupo')
                        valuesList.append('cab__cab__grupo')
                        if grupo != 0:
                            filtro = filtro & (Q(cab__cab__grupo=grupo))
                    if porpais:
                        listDistinct.append('cab__pais__nombre')
                        listOrderBy.append('cab__pais__nombre')
                        valuesList.append('cab__pais__nombre')
                        valuesList.append('cab__pais__id')
                        if pais != '0':
                            filtro = filtro & (Q(cab__pais__nombre__icontains=pais))
                    if poruniversidad:
                        listDistinct.append('cab__universidad__nombre')
                        listOrderBy.append('cab__universidad__nombre')
                        valuesList.append('cab__universidad__nombre')
                        if universidad != '0':
                            filtro = filtro & (Q(cab__universidad__nombre__icontains=universidad))
                    if porgenero:
                        listDistinct.append('cab__genero')
                        listOrderBy.append('cab__genero')
                        valuesList.append('cab__genero')
                    if criterio:
                        search = criterio.strip()
                        filtro = filtro & (Q(indicador__icontains=search) | Q(cab__pais__nombre__icontains=search) | Q(cab__universidad__nombre__icontains=search) | Q(cab__otrauniversidad__icontains=search) | Q(pregunta__indicador__factores__nombres__icontains=search))
                    listDistinct.append('indicador')
                    listOrderBy.append('indicador')
                    valuesList.append('indicador')
                    valuesList.append('pregunta__indicador__id')
                    valuesList.append('pregunta__indicador__factores__nombres')
                    valuesList.append('cab__cab__id')
                    listCount = GEDCRespuestas.objects.select_related('cab').filter(cab__cab__publicar=True).filter(cab__pais__isnull=False).filter(filtro).distinct(*listDistinct).values(*valuesList).count()
                    rows = GEDCRespuestas.objects.select_related('cab').filter(cab__cab__publicar=True).filter(cab__pais__isnull=False).filter(filtro).distinct(*listDistinct).values(*valuesList).order_by(*listOrderBy)[desde:hasta]
                    for row in rows:
                        grupoid = row['cab__cab__grupo'] if 'cab__cab__grupo' in row else None
                        grupo_nombre = dict(GEDC_GRUPO)[grupoid] if grupoid else ''
                        paisid = row['cab__pais__id'] if 'cab__pais__id' in row else None
                        pais_nombre = row['cab__pais__nombre'] if 'cab__pais__nombre' in row else ''
                        universidad_nombre = row['cab__universidad__nombre'] if 'cab__universidad__nombre' in row else ''
                        generoid = row['cab__genero'] if 'cab__genero' in row else None
                        genero_nombre = dict(GENEROS_ENCUESTAS)[generoid] if generoid else ''
                        preguntaid = row['pregunta__indicador__id'] if 'pregunta__indicador__id' in row else None
                        pregunta_nombre = row['indicador'] if 'indicador' in row else ''
                        factor_nombre = row['pregunta__indicador__factores__nombres'] if 'pregunta__indicador__factores__nombres' in row else ''
                        frespuesta = Q(status=True) & Q(respcalificacion_inversa__isnull=False)
                        if grupoid:
                            frespuesta = frespuesta & Q(cab__cab__grupo=grupoid)
                        if paisid:
                            frespuesta = frespuesta & Q(cab__pais__id=paisid)
                        if universidad_nombre:
                            frespuesta = frespuesta & Q(cab__universidad__nombre=universidad_nombre)
                        if generoid:
                            frespuesta = frespuesta & Q(cab__genero=generoid)
                        if preguntaid:
                            frespuesta = frespuesta & Q(pregunta__indicador__id=preguntaid)
                        listaRespuestas = GEDCRespuestas.objects.filter(cab__cab__publicar=True).filter(frespuesta).values_list('respcalificacion_inversa', flat=True)
                        totpreguntas = listaRespuestas.count()
                        media = round(stats.mean(list(listaRespuestas)),2)
                        desvestandar = round(stats.pstdev(list(listaRespuestas)),2)
                        aaData.append([grupo_nombre, pais_nombre, universidad_nombre, genero_nombre, factor_nombre, pregunta_nombre, totpreguntas, media, desvestandar, porgrupo, porpais, poruniversidad, porgenero])
                return JsonResponse({"result": "ok", "data": aaData, "iTotalRecords": listCount, "iTotalDisplayRecords": listCount})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al cargar los datos. %s" % ex.__str__(), "data": [], "iTotalRecords": 0, "iTotalDisplayRecords": 0})

    else:
        if 'action' in request.GET:
            data['action'] = action = request.GET['action']

            if action == 'consultaruniversidades':
                try:
                    lista = []
                    id = request.GET['id']
                    filtro = InstitucionEducacionSuperior.objects.filter(status=True, pais__nombre=id).distinct('nombre').order_by('nombre')
                    lista = ["{}".format(cr.nombre) for cr in filtro]
                    return JsonResponse({'result': 'ok', 'lista': lista})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            if action == 'listResultadosGrafica':
                try:
                    grupo, pais, universidad, variable, calculo = int(request.GET.get('grupo','')), request.GET.get('pais',''), request.GET.get('universidad',''), request.GET.get('variable',''), request.GET.get('calculo',''),
                    porgrupo, porpais, poruniversidad, porgenero = False, False, False, False
                    if variable:
                        listavariables = variable.split(',')
                        for lv in listavariables:
                            if int(lv) == 1:
                                porgrupo = True
                            if int(lv) == 2:
                                porpais = True
                            if int(lv) == 3:
                                poruniversidad = True
                            if int(lv) == 4:
                                porgenero = True
                    qsbase = GEDCPersona.objects.select_related('cab').filter(status=True, cab__publicar=True, pais__isnull=False, universidad__isnull=False)
                    filtro, listDistinct, listOrderBy, valuesList = Q(status=True), [], [], []
                    if porgrupo:
                        listDistinct.append('cab__grupo')
                        listOrderBy.append('cab__grupo')
                        valuesList.append('cab__grupo')
                    if grupo != 0:
                        filtro = filtro & (Q(cab__grupo=grupo))
                    if porpais:
                        listDistinct.append('pais__nombre')
                        listOrderBy.append('pais__nombre')
                        valuesList.append('pais__nombre')
                        valuesList.append('pais__id')
                    if pais != '0':
                        filtro = filtro & (Q(pais__nombre=pais))
                    if poruniversidad:
                        listDistinct.append('universidad__nombre')
                        listOrderBy.append('universidad__nombre')
                        valuesList.append('universidad__nombre')
                    if universidad != '0':
                        filtro = filtro & Q(universidad__nombre=universidad)
                    if porgenero:
                        listDistinct.append('genero')
                        listOrderBy.append('genero')
                        valuesList.append('genero')
                    rows = qsbase.filter(filtro).distinct(*listDistinct).values(*valuesList).order_by(*listOrderBy)
                    listCount = rows.count()
                    data['porgrupo'] = porgrupo
                    data['porpais'] = porpais
                    data['poruniversidad'] = poruniversidad
                    data['porgenero'] = porgenero
                    data['filtro'] = filtro
                    data['listDistinct'] = listDistinct
                    data['listOrderBy'] = listOrderBy
                    data['valuesList'] = valuesList
                    data['listadata'] = rows
                    data['listcount'] = listCount
                    data['calculo'] = int(calculo)
                    factores_list = GEDCIndicador.objects.filter(status=True, factores__isnull=False).distinct('factores').values_list('factores__pk', flat=True)
                    data['factores'] = GEDCFactores.objects.filter(status=True, pk__in=factores_list).order_by('nombres')
                    template = get_template("adm_gedcevaluacion/graficaEstadistica.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'listaResultadosPaginacion':
                try:
                    criterio, grupo, pais, universidad, variable, calculo = request.GET.get('criterio',''), int(request.GET.get('grupo','')), request.GET.get('pais',''), request.GET.get('universidad',''), request.GET.get('variable',''), request.GET.get('calculo',''),
                    porgrupo, porpais, poruniversidad, porgenero = False, False, False, False
                    if variable:
                        listavariables = variable.split(',')
                        for lv in listavariables:
                            if int(lv) == 1:
                                porgrupo = True
                            if int(lv) == 2:
                                porpais = True
                            if int(lv) == 3:
                                poruniversidad = True
                            if int(lv) == 4:
                                porgenero = True
                    aaData = []
                    filtro, listDistinct, listOrderBy, valuesList = Q(status=True), [], [], []
                    if porgrupo:
                        listDistinct.append('cab__cab__grupo')
                        listOrderBy.append('cab__cab__grupo')
                        valuesList.append('cab__cab__grupo')
                    if grupo != 0:
                        filtro = filtro & (Q(cab__cab__grupo=grupo))
                    if porpais:
                        listDistinct.append('cab__pais__nombre')
                        listOrderBy.append('cab__pais__nombre')
                        valuesList.append('cab__pais__nombre')
                        valuesList.append('cab__pais__id')
                    if pais != '0':
                        filtro = filtro & (Q(cab__pais__nombre=pais))
                    if poruniversidad:
                        listDistinct.append('cab__universidad__nombre')
                        listOrderBy.append('cab__universidad__nombre')
                        valuesList.append('cab__universidad__nombre')
                    if universidad != '0':
                        filtro = filtro & Q(cab__universidad__nombre=universidad)
                    if porgenero:
                        listDistinct.append('cab__genero')
                        listOrderBy.append('cab__genero')
                        valuesList.append('cab__genero')
                    if criterio:
                        search = criterio.strip()
                        filtro = filtro & (Q(indicador__indicador__icontains=search) | Q(indicador__factores__nombres__icontains=search))
                        if poruniversidad:
                            filtro = filtro | Q(cab__universidad__nombre__icontains=search)
                        if porpais:
                            filtro = filtro | Q(cab__pais__nombre__icontains=search)
                    listDistinct.append('indicador')
                    listOrderBy.append('indicador')
                    valuesList.append('indicador__indicador')
                    valuesList.append('indicador__id')
                    valuesList.append('indicador__factores__nombres')
                    # valuesList.append('cab__cab__id')
                    rows = GEDCRespuestas.objects.select_related('cab').filter(cab__cab__publicar=True).filter(cab__pais__isnull=False, cab__universidad__isnull=False).filter(filtro).distinct(*listDistinct).values(*valuesList).order_by(*listOrderBy)
                    listCount = rows.count()
                    data['porgrupo'] = porgrupo
                    data['porpais'] = porpais
                    data['poruniversidad'] = poruniversidad
                    data['porgenero'] = porgenero
                    data['listCount'] = listCount
                    data['filtro'] = filtro
                    # for row in rows:
                    #     grupoid = row['cab__cab__grupo'] if 'cab__cab__grupo' in row else None
                    #     grupo_nombre = dict(GEDC_GRUPO)[grupoid] if grupoid else ''
                    #     paisid = row['cab__pais__id'] if 'cab__pais__id' in row else None
                    #     pais_nombre = row['cab__pais__nombre'] if 'cab__pais__nombre' in row else ''
                    #     universidad_nombre = row['cab__universidad__nombre'] if 'cab__universidad__nombre' in row else ''
                    #     generoid = row['cab__genero'] if 'cab__genero' in row else None
                    #     genero_nombre = dict(GENEROS_ENCUESTAS)[generoid] if generoid else ''
                    #     preguntaid = row['pregunta__indicador__id'] if 'pregunta__indicador__id' in row else None
                    #     pregunta_nombre = row['indicador'] if 'indicador' in row else ''
                    #     factor_nombre = row['pregunta__indicador__factores__nombres'] if 'pregunta__indicador__factores__nombres' in row else ''
                    #     frespuesta = Q(status=True) & Q(respcalificacion_inversa__isnull=False)
                    #     if grupoid:
                    #         frespuesta = frespuesta & Q(cab__cab__grupo=grupoid)
                    #     if paisid:
                    #         frespuesta = frespuesta & Q(cab__pais__id=paisid)
                    #     if universidad_nombre:
                    #         frespuesta = frespuesta & Q(cab__universidad__nombre=universidad_nombre)
                    #     if generoid:
                    #         frespuesta = frespuesta & Q(cab__genero=generoid)
                    #     if preguntaid:
                    #         frespuesta = frespuesta & Q(pregunta__indicador__id=preguntaid)
                    #     listaRespuestas = GEDCRespuestas.objects.filter(cab__cab__publicar=True).filter(frespuesta).values_list('respcalificacion_inversa', flat=True)
                    #     totpreguntas = listaRespuestas.count()
                    #     media = round(stats.mean(list(listaRespuestas)),2)
                    #     desvestandar = round(stats.pstdev(list(listaRespuestas)),2)
                    #     aaData.append([grupo_nombre, pais_nombre, universidad_nombre, genero_nombre, factor_nombre, pregunta_nombre, totpreguntas, media, desvestandar, porgrupo, porpais, poruniversidad, porgenero])
                    paging = MiPaginador(rows, 15)
                    p = 1
                    try:
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        page = paging.page(p)
                    except:
                        page = paging.page(1)
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['listado'] = page.object_list
                    template = get_template("adm_gedcevaluacion/paginacion.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

            if action == 'addencuesta':
                try:
                    data['title'] = u'Adicionar Encuesta'
                    data['form'] = GEDCCabForm()
                    return render(request, "adm_gedcevaluacion/addencuesta.html", data)
                except Exception as ex:
                    pass

            if action == 'editencuesta':
                try:
                    data['title'] = u'Editar Encuesta'
                    data['filtro'] = filtro = GEDCCab.objects.get(pk=int(request.GET['id']))
                    data['form'] = GEDCCabForm(initial=model_to_dict(filtro))
                    return render(request, "adm_gedcevaluacion/editencuesta.html", data)
                except Exception as ex:
                    pass

            if action == 'delencuesta':
                try:
                    data['title'] = u'ELIMINAR ENCUESTA'
                    data['filtro'] = GEDCCab.objects.get(pk=request.GET['id'])
                    return render(request, 'adm_gedcevaluacion/delencuesta.html', data)
                except Exception as ex:
                    pass

            if action == 'delpregunta':
                try:
                    data['title'] = u'ELIMINAR PREGUNTA'
                    data['filtro'] = GEDCPreguntas.objects.get(pk=request.GET['id'])
                    return render(request, 'adm_gedcevaluacion/delpregunta.html', data)
                except Exception as ex:
                    pass

            if action == 'preguntas':
                data['title'] = u'Preguntas'
                data['cab'] = cab = GEDCCab.objects.get(pk=request.GET['id'])
                data['titulo'] = 'Preguntas {}'.format(cab.nombre)
                data['preguntas'] = preguntas = GEDCPreguntas.objects.filter(cab=cab, status=True).order_by('orden')
                return render(request, 'adm_gedcevaluacion/preguntas.html', data)

            if action == 'addpregunta':
                try:
                    data['id'] = id = int(request.GET['id'])
                    filtro = GEDCCab.objects.get(pk=id)
                    query = GEDCIndicador.objects.filter(status=True, grupo=filtro.grupo)
                    form = GEDCPreguntasForm()
                    form.fields['orden'].initial = 1
                    if GEDCPreguntas.objects.filter(cab_id=id).exists():
                        preguntalast = GEDCPreguntas.objects.filter(cab_id=id).order_by('orden').last()
                        query = query.exclude(pk__in=GEDCPreguntas.objects.filter(cab_id=id).values_list('indicador__pk',flat=True))
                        form.fields['orden'].initial = preguntalast.orden + 1
                    form.fields['indicador'].queryset = query.order_by('pk')
                    data['form'] = form
                    template = get_template("adm_gedcevaluacion/modal/formpregunta.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'editpregunta':
                try:
                    data['id'] = id = int(request.GET['id'])
                    filtro = GEDCPreguntas.objects.get(pk=id)
                    form = GEDCPreguntasEditarForm(initial=model_to_dict(filtro))
                    data['form'] = form
                    template = get_template("adm_gedcevaluacion/modal/formpregunta.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'consultarindicador':
                try:
                    id = request.GET['id']
                    filtro = GEDCIndicador.objects.get(pk=int(id))
                    dict_filtro = model_to_dict(filtro)
                    response = JsonResponse({'state': True, 'modelo': dict_filtro})
                except Exception as ex:
                    response = JsonResponse({'state': False, 'requerido': False})
                return HttpResponse(response.content)

            if action == 'configuraciones':
                try:
                    data['title'] = u'Configuraciones'
                    data['indicadores'] = GEDCIndicador.objects.filter(status=True).order_by('id')
                    data['factores'] = GEDCFactores.objects.filter(status=True).order_by('nombres')
                    return render(request, "adm_gedcevaluacion/configuraciones.html", data)
                except Exception as ex:
                    pass

            if action == 'estadistica':
                try:
                    data['title'] = u'Gedc Estadistica'
                    data['tp_calculo_gedc'] = TIPO_CALCULO_GEDC
                    data['tp_detalle_gedc'] = GEDC_DETALLE_BUSQUEDA
                    data['cab'] = GEDCCab.objects.filter(status=True, publicar=True).order_by('id')
                    return render(request, "adm_gedcevaluacion/reportes.html", data)
                except Exception as ex:
                    pass

            if action == 'addindicador':
                try:
                    data['form2'] = GEDCIndicadoresForm()
                    template = get_template("adm_gedcevaluacion/modal/formindicador.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'editindicador':
                try:
                    data['id'] = request.GET['id']
                    data['filtro'] = filtro = GEDCIndicador.objects.get(pk=request.GET['id'])
                    data['form2'] = GEDCIndicadoresForm(initial=model_to_dict(filtro))
                    template = get_template("adm_gedcevaluacion/modal/formindicador.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'delindicador':
                try:
                    data['title'] = u'ELIMINAR PREGUNTA'
                    data['categoria'] = GEDCIndicador.objects.get(pk=request.GET['id'])
                    return render(request, 'adm_gedcevaluacion/delindicador.html', data)
                except Exception as ex:
                    pass

            if action == 'addfactor':
                try:
                    data['form2'] = GEDCFactorForm()
                    template = get_template("adm_gedcevaluacion/modal/formindicador.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'editfactor':
                try:
                    data['id'] = request.GET['id']
                    data['filtro'] = filtro = GEDCFactores.objects.get(pk=request.GET['id'])
                    data['form2'] = GEDCFactorForm(initial=model_to_dict(filtro))
                    template = get_template("adm_gedcevaluacion/modal/formindicador.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'delfactor':
                try:
                    data['title'] = u'ELIMINAR FACTOR'
                    data['categoria'] = GEDCFactores.objects.get(pk=request.GET['id'])
                    return render(request, 'adm_gedcevaluacion/delindicador.html', data)
                except Exception as ex:
                    pass

            if action == 'encuestados':
                data['cab'] = cab = GEDCCab.objects.get(pk=request.GET['id'])
                data['title'] = 'Evaluados'
                url_vars = ''
                url_vars += '&action={}&id={}'.format(action, request.GET['id'])
                filtro = Q(status=True)
                listado = GEDCPersona.objects.filter(cab=cab).filter(filtro).order_by('pk')
                paging = MiPaginador(listado, 20)
                p = 1
                try:
                    paginasesion = 1
                    if 'paginador' in request.session:
                        paginasesion = int(request.session['paginador'])
                    if 'page' in request.GET:
                        p = int(request.GET['page'])
                        p = int(request.GET['page'])
                    else:
                        p = paginasesion
                    try:
                        page = paging.page(p)
                    except:
                        p = 1
                    page = paging.page(p)
                except:
                    page = paging.page(p)
                request.session['paginador'] = p
                data['paging'] = paging
                data['rangospaging'] = paging.rangos_paginado(p)
                data['page'] = page
                data['listado'] = page.object_list
                data["url_vars"] = url_vars
                return render(request, 'adm_gedcevaluacion/encuestados.html', data)


            if action == 'generarexcel':
                pk = request.GET['id']
                cab = GEDCCab.objects.get(pk=request.GET['id'])
                __author__ = 'Unemi'
                borders = Borders()
                borders.left = Borders.THIN
                borders.right = Borders.THIN
                borders.top = Borders.THIN
                borders.bottom = Borders.THIN
                align = Alignment()
                font_style = XFStyle()
                font_style.font.bold = True
                font_style.borders = borders
                font_style.alignment = align
                font_style2 = XFStyle()
                font_style2.font.bold = False
                font_style2.borders = borders
                font_style2.alignment = align
                style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
                style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                style1 = easyxf(num_format_str='D-MMM-YY')
                title = easyxf('font: name Calibri, color-index black, bold on , height 260; alignment: horiz centre')
                fuentecabecera = easyxf('font: name Calibri, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: horiz centre; borders: left thin, right thin, top thin, bottom thin')
                fuentenormal = easyxf('font: name Calibri, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
                fuentetexto = easyxf('font: name Calibri, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
                font_style = XFStyle()
                font_style.font.bold = True
                font_style2 = XFStyle()
                font_style2.font.bold = False
                wb = Workbook(encoding='utf-8')
                ws = wb.add_sheet('consolidado')
                # ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                response = HttpResponse(content_type="application/ms-excel")
                response['Content-Disposition'] = 'attachment; filename=gedc_respuestas_' + random.randint(1, 10000).__str__() + '.xls'
                ws.write_merge(0, 3, 0, 3, 'RESPUESTA ENCUESTA {}'.format(cab.nombre), title)
                columns = [
                    (u"MARCA TEMPORAL", 4000),
                    (u"PAÍS", 6000),
                    (u"OTRO PAIS", 6000),
                    (u"UNIVERSIDAD", 4000),
                    (u"OTRA UNIVERSIDAD", 6000),
                    (u"GENERO UNIVERSIDAD", 6000),
                ]
                for cp in cab.gedcpreguntas_set.filter(status=True).order_by('orden'):
                    columns.append((cp.indicador.indicador, 10000))
                row_num = 5
                for col_num in range(len(columns)):
                    ws.write(row_num, col_num, columns[col_num][0], fuentecabecera)
                    ws.col(col_num).width = columns[col_num][1]
                row_num = 6
                for q in cab.gedcpersona_set.filter(status=True):
                    ws.write_merge(row_num, row_num, 0, 0, str(q.fecha_creacion), font_style2)
                    if q.pais:
                        ws.write_merge(row_num, row_num, 1, 1, q.pais.nombre, font_style2)
                    ws.write_merge(row_num, row_num, 2, 2, q.otropais, font_style2)
                    if q.universidad:
                        ws.write_merge(row_num, row_num, 3, 3, q.universidad.nombre, font_style2)
                    ws.write_merge(row_num, row_num, 4, 4, q.otrauniversidad, font_style2)
                    ws.write_merge(row_num, row_num, 5, 5, q.get_genero(), font_style2)
                    col_row = 6
                    for qi in q.respuestas():
                        if  qi.pregunta.indicador.evalua:
                            ws.write_merge(row_num, row_num, col_row, col_row, qi.get_respevalua(), font_style2)
                        if  qi.pregunta.indicador.calificacion:
                            ws.write_merge(row_num, row_num, col_row, col_row, qi.respcalificacion_inversa, font_style2)
                        if  qi.pregunta.indicador.observacion:
                            ws.write_merge(row_num, row_num, col_row, col_row, qi.respobservacion, font_style2)
                        col_row += 1
                    row_num += 1

                date_format = xlwt.XFStyle()
                date_format.num_format_str = 'yyyy-mm-dd'
                date_formatreverse = xlwt.XFStyle()
                date_formatreverse.num_format_str = 'dd/mm/yyyy'
                wb.save(response)
                return response

        else:
            data['title'] = u'Encuestas GEDC'
            url_vars = ''
            filtro = Q(status=True)
            ids = None
            data['search'] = search = request.GET.get('s', '')

            if search:
                filtro = filtro & (Q(nombre__icontains=search) | (Q(departamento__nombre__icontains=search)))
                url_vars += '&s=' + search

            procesos = GEDCCab.objects.filter(filtro).order_by('-id')

            paging = MiPaginador(procesos, 20)
            p = 1
            try:
                paginasesion = 1
                if 'paginador' in request.session:
                    paginasesion = int(request.session['paginador'])
                if 'page' in request.GET:
                    p = int(request.GET['page'])
                    p = int(request.GET['page'])
                else:
                    p = paginasesion
                try:
                    page = paging.page(p)
                except:
                    p = 1
                page = paging.page(p)
            except:
                page = paging.page(p)
            request.session['paginador'] = p
            data['paging'] = paging
            data['rangospaging'] = paging.rangos_paginado(p)
            data['page'] = page
            data["url_vars"] = url_vars
            data['ids'] = ids if ids else ""
            data['proceso'] = page.object_list
            data['email_domain'] = EMAIL_DOMAIN
            return render(request, 'adm_gedcevaluacion/view.html', data)