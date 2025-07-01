# -*- coding: UTF-8 -*-
import sys

import openpyxl
import xlrd as xlrd
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import Group, User
from django.core.exceptions import ObjectDoesNotExist
from django.db import transaction, connection, connections
from django.db.models import Sum, Count, Exists, OuterRef, Subquery
from django.db.models.functions import Coalesce
from django.forms import model_to_dict
from xlwt import Workbook
import xlwt
import random
from xlwt import *
from django.template.context import Context
from django.template.loader import get_template
from django.db.models.query_utils import Q
from django.http import HttpResponseRedirect, JsonResponse, HttpResponse
from django.shortcuts import render
import json

from bd.models import LogEntryLogin
from inno.models import FechaPlanificacionSedeVirtualExamen, TurnoPlanificacionSedeVirtualExamen, \
    AulaPlanificacionSedeVirtualExamen, MateriaAsignadaPlanificacionSedeVirtualExamen
from mobile.views import make_thumb_fotopersona, make_thumb_picture
from decorators import secure_module, last_access
from sga.commonviews import adduserdata, obtener_reporte, ficha_socioeconomica, traerNotificaciones
from sga.excelbackground import actualizar_visible_horario_masivo, limpiar_cache_masivo, reporte_auditoria_background, \
    reporte_titulacionporcarrera
from sga.forms import InscripcionForm, RecordAcademicoForm, HistoricoRecordAcademicoForm, CargarFotoForm, \
    DocumentoInscripcionForm, ConvalidacionInscripcionForm, \
    CambiomallaForm, HomologacionInscripcionForm, ConsiderarForm, NuevaInscripcionForm, CambioGrupoForm, \
    InscripcionTipoInscripcionForm, CambionivelmallaForm, \
    FechaInicioConvalidacionInscripcionForm, RetiradoMateriaForm, ImportarArchivoXLSForm, \
    FechaInicioPrimerNivelInscripcionForm, FechaMateriaRecordForm, PersonaTituloUniversidadForm, \
    FechaInicioCarreraInscripcionForm, AutorizarAlumnoSolicitudForm, MallaHistoricaForm, NombreUsuarioForm, \
    PersonaPPLForm, MatriculaSedeExamenForm
from settings import DEFAULT_PASSWORD, ALUMNOS_GROUP_ID, UTILIZA_GRUPOS_ALUMNOS, EMAIL_DOMAIN, ARCHIVO_TIPO_GENERAL, EMAIL_INSTITUCIONAL_AUTOMATICO, \
    PREGUNTAS_INSCRIPCION, CORREO_OBLIGATORIO, GENERAR_TUMBAIL, \
    USA_TIPOS_INSCRIPCIONES, TIPO_INSCRIPCION_INICIAL, CONTROL_UNICO_CREDENCIALES, \
    MATRICULACION_LIBRE, PROFESORES_GROUP_ID, NOTA_ESTADO_APROBADO, ESTADO_GESTACION
from sga.funciones import log, lista_correo, calculate_username, generar_usuario, \
    generar_nombre, resetear_clave, puede_realizar_accion, puede_modificar_inscripcion, MiPaginador, variable_valor, \
    convertir_fecha, resetear_clave_pregrado_virtual, null_to_decimal, puede_realizar_accion_afirmativo, null_to_numeric
from sga.funciones_templatepdf import CHOICES_FUNCION_VIEW_REQUISITO
from sga.models import Persona, PersonaReligion, PersonaDocumentoPersonal, Inscripcion, ExamenComplexivo, \
    RecordAcademico, DocumentosDeInscripcion, \
    HistoricoRecordAcademico, FotoPersona, Archivo, Grupo, ConvalidacionInscripcion, NivelMalla, EjeFormativo, \
    AsignaturaMalla, Carrera, SeguimientoEstudiante, miinstitucion, PreInscrito, InscripcionTipoInscripcion, \
    Asignatura, Nivel, Matricula, Clase, Sesion, RetiroCarrera, Modalidad, Sede, Administrativo, Profesor, \
    TiempoDedicacionDocente, Coordinacion, PersonaTituloUniversidad, CUENTAS_CORREOS, \
    PracticasPreprofesionalesInscripcion, AutorizarAlumnoSolicitud, PaeInscripcionActividades, PaeFechaActividad, \
    VirtualSoporteUsuario, VirtualSoporteUsuarioInscripcion, EncuestaTecnologica, PreguntaEncuestaTecnologica, \
    TarjetaRegistroAcademico, AgregacionEliminacionMaterias, HistorialPersonaPPL, InscripcionMalla, \
    HomologacionInscripcion, InscripcionTesDrive, LogEntryBackup, LogEntryBackupdos, MateriaAsignada, \
    MatriculaTitulacion, ArchivoTitulacion, \
    ComplexivoExamenDetalle, ComplexivoExamen, ComplexivoGrupoTematica, TIPO_ARCHIVO_COMPLEXIVO_PROPUESTA, \
    ComplexivoDetalleGrupo, RequisitoSustentar, \
    AlternativaTitulacion, TIPO_CELULAR, ModuloMalla, ParticipantesMatrices, ComplexivoAcompanamiento, \
    ComplexivoTematica, PerdidaGratuidad, DetalleModeloEvaluativo, SedeVirtual, Notificacion, Graduado, Egresado, \
    PerfilUsuario, Malla, Materia,MateriaTitulacion
from inno.funciones import estar_matriculado_todas_asignaturas_ultimo_periodo_academico, \
    asignaturas_aprobadas_primero_penultimo_nivel, ficha_estudiantil_actualizada_completa, \
    haber_aprobado_modulos_ingles, haber_aprobado_modulos_computacion, \
    haber_cumplido_horas_creditos_practicas_preprofesionales, haber_cumplido_horas_creditos_vinculacion, \
    no_adeudar_institucion, asignaturas_aprobadas_primero_ultimo_nivel
from sga.tasks import send_html_mail, conectar_cuenta
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
from sga.funcionesxhtml2pdf import conviert_html_to_pdf
from sga.templatetags.sga_extras import encrypt
from django.contrib.admin.models import LogEntry, ADDITION, DELETION

unicode =str

@login_required(redirect_field_name='ret', login_url='/loginsga')
@secure_module
@last_access
@transaction.atomic()
def view(request):
    data = {}
    adduserdata(request, data)
    data['personasesion'] = personasesion = request.session['persona']
    periodo = request.session['periodo']
    if request.method == 'POST':
        action = request.POST['action']

        if action == 'cambiomalla':
            try:
                inscripcion = Inscripcion.objects.get(pk=request.POST['id'])
                f = CambiomallaForm(request.POST)
                if f.is_valid():
                    # malla = inscripcion.inscripcionmalla_set.all()
                    # malla.delete()
                    for eInscripcionMalla in inscripcion.inscripcionmalla_set.all():
                        eInscripcionMalla.status = False
                        eInscripcionMalla.save(request)
                    eInscripcionMallas = inscripcion.inscripcionmalla_set.filter(inscripcion=inscripcion, malla=f.cleaned_data['malla_nueva'])
                    if eInscripcionMallas.values("id").exists():
                        eInscripcionMalla = eInscripcionMallas[0]
                        eInscripcionMalla.status=True
                    else:
                        eInscripcionMalla = InscripcionMalla(inscripcion=inscripcion,
                                                             malla=f.cleaned_data['malla_nueva'])
                    eInscripcionMalla.save(request)
                    inscripcion.actualizar_creditos()
                    inscripcion.actualizar_nivel()
                    log(u'Modifico malla de inscripcion: %s - %s' % (inscripcion.persona, eInscripcionMalla.malla), request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'cambionivel':
            try:
                inscripcion = Inscripcion.objects.get(pk=request.POST['id'])
                f = CambionivelmallaForm(request.POST)
                if f.is_valid():
                    nivel = inscripcion.mi_nivel()
                    nivel.nivel = f.cleaned_data['nuevonivel']
                    nivel.save(request)
                    log(u'Modifico nivel de inscripcion: %s - %s' % (inscripcion.persona, nivel.nivel), request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'add_titulos_profesionales':
            try:
                f = PersonaTituloUniversidadForm(request.POST)
                if f.is_valid():
                    persona = Persona.objects.get(pk=request.POST['idp'])
                    personatitulouniversidad = PersonaTituloUniversidad(persona=persona,
                                                                        codigoregistro=f.cleaned_data['codigoregistro'],
                                                                        fecharegistro=f.cleaned_data['fecharegistro'],
                                                                        fechaacta=f.cleaned_data['fechaacta'],
                                                                        fechainicio=f.cleaned_data['fechainicio'],
                                                                        fecharegresado=f.cleaned_data['fecharegresado'],
                                                                        universidad=f.cleaned_data['universidad'],
                                                                        tipouniversidad=f.cleaned_data['tipouniversidad'],
                                                                        tiponivel=f.cleaned_data['tiponivel'],
                                                                        nombrecarrera=f.cleaned_data['nombrecarrera'])
                    personatitulouniversidad.save(request)
                    log(u'Agrego titulo profesional: %s' % personatitulouniversidad, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'edit_titulos_profesionales':
            try:
                f = PersonaTituloUniversidadForm(request.POST)
                if f.is_valid():
                    personatitulouniversidad = PersonaTituloUniversidad.objects.get(pk=request.POST['id'])
                    personatitulouniversidad.codigoregistro = f.cleaned_data['codigoregistro']
                    personatitulouniversidad.fecharegistro = f.cleaned_data['fecharegistro']
                    personatitulouniversidad.fechaacta = f.cleaned_data['fechaacta']
                    personatitulouniversidad.fechainicio = f.cleaned_data['fechainicio']
                    personatitulouniversidad.fecharegresado = f.cleaned_data['fecharegresado']
                    personatitulouniversidad.universidad = f.cleaned_data['universidad']
                    personatitulouniversidad.tipouniversidad = f.cleaned_data['tipouniversidad']
                    personatitulouniversidad.tiponivel = f.cleaned_data['tiponivel']
                    personatitulouniversidad.nombrecarrera = f.cleaned_data['nombrecarrera']
                    personatitulouniversidad.save(request)
                    log(u'Edito titulo profesional: %s' % personatitulouniversidad, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'delete_titulos_profesionales':
            try:
                personatitulouniversidad = PersonaTituloUniversidad.objects.get(pk=request.POST['id'])
                personatitulouniversidad.delete()
                log(u'Elimino titulo profesional: %s' % personatitulouniversidad, request, "del")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'deleteppl':
            try:
                puede_realizar_accion(request, 'sga.puede_eliminar_ppl')
                hppl = HistorialPersonaPPL.objects.get(pk=request.POST['id'])
                hppl.delete()
                log(u'Elimino registro de ppl: %s' % hppl, request, "del")
                messages.add_message(request, messages.SUCCESS, f'Se elimino correctamente el registro')
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'editfechamateria':
            try:
                recordacademico = RecordAcademico.objects.get(pk=request.POST['id'])
                f = FechaMateriaRecordForm(request.POST)
                if f.is_valid():
                    recordacademico.fechainicio = f.cleaned_data['inicio']
                    recordacademico.fechafin = f.cleaned_data['fin']
                    recordacademico.save(request)
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'pdflistaactividades':
            try:
                data = {}
                data['fechaactual'] = datetime.now()
                data['personasesion'] = personasesion
                data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.POST['id'])
                data['inscrito'] = PaeInscripcionActividades.objects.select_related().filter(matricula__inscripcion=inscripcion, status=True, matricula__status=True).order_by('matricula__nivel__periodo__id')
                data['decano'] = inscripcion.coordinacion.responsable_periododos(periodo,1)
                return conviert_html_to_pdf(
                    'inscripciones/listaactividades_pdf.html',
                    {
                        'pagesize': 'A4',
                        'data': data,
                    }
                )
            except Exception as ex:
                pass

        elif action == 'cambiocategoria':
            try:
                inscripcion = Inscripcion.objects.get(pk=request.POST['id'])
                f = InscripcionTipoInscripcionForm(request.POST)
                if f.is_valid():
                    if inscripcion.tipo_inscripcion():
                        tipo = inscripcion.tipo_inscripcion()
                        tipo.tipoinscripcion = f.cleaned_data['tipoinscripcion']
                        tipo.save(request)
                    else:
                        tipo = InscripcionTipoInscripcion(inscripcion=inscripcion,
                                                          tipoinscripcion=f.cleaned_data['tipoinscripcion'])
                        tipo.save(request)
                    log(u'Modifico categoria de inscripcion: %s' % inscripcion.persona, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'chequeoinscripcionedit':
            try:
                cedula = request.POST['cedula']
                carrera = request.POST['carrera']
                idi = request.POST['id']
                if cedula.__len__() > 0 and carrera.__len__() > 0:
                    if not Inscripcion.objects.filter(persona__cedula=request.POST['cedula'], carrera=Carrera.objects.get(pk=carrera)):
                        return JsonResponse({"result": "ok"})
                    else:
                        ins = Inscripcion.objects.filter(persona__cedula=request.POST['cedula'], carrera=Carrera.objects.get(pk=carrera))[0]
                        if ins.id != idi:
                            return JsonResponse({"result": "bad"})
                        else:
                            return JsonResponse({"result": "ok"})
                else:
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                return JsonResponse({"result": "bad"})

        elif action == 'addrecord':
            try:
                inscripcion = Inscripcion.objects.get(pk=request.POST['id'])
                f = RecordAcademicoForm(request.POST)
                if f.is_valid():
                    if f.cleaned_data['asignatura']:
                        if inscripcion.recordacademico_set.filter(asignatura=f.cleaned_data['asignatura']).exists():
                            return JsonResponse({"result": "bad", "mensaje": u"Ya existe la asignatura en el record, modifiquela desde el historico."})
                    record = RecordAcademico(inscripcion=inscripcion,
                                             asignatura=f.cleaned_data['asignatura'],
                                             modulomalla=inscripcion.asignatura_en_modulomalla(f.cleaned_data['asignatura']),
                                             # asignaturamalla=inscripcion.asignatura_en_asignaturamalla(f.cleaned_data['asignatura'] if f.cleaned_data['asignatura'] else f.cleaned_data['asignaturamallahistorico'].asignatura ),
                                             nota=f.cleaned_data['nota'],
                                             asistencia=f.cleaned_data['asistencia'],
                                             fecha=f.cleaned_data['fecha'],
                                             aprobada=f.cleaned_data['aprobada'],
                                             noaplica=f.cleaned_data['noaplica'],
                                             convalidacion=f.cleaned_data['convalidacion'],
                                             pendiente=False,
                                             creditos=f.cleaned_data['creditos'],
                                             horas=f.cleaned_data['horas'],
                                             homologada=f.cleaned_data['homologada'],
                                             valida=f.cleaned_data['valida'],
                                             validapromedio=f.cleaned_data['validapromedio'],
                                             observaciones= f.cleaned_data['observaciones'],
                                             suficiencia=f.cleaned_data['suficiencia'])
                    # asignaturamallahistorico=f.cleaned_data['asignaturamallahistorico'])
                    record.save(request)
                    record.actualizar()
                    inscripcion.actualizar_nivel()
                    inscripcion.actualiza_matriculas(record.asignatura if record.asignatura else record.asignaturamallahistorico.asignatura)
                    log(u'Adiciono record academico: %s - %s' % (record, record.inscripcion.persona), request, "add")

                    lista_email_cco = ['sga@unemi.edu.ec']
                    tituloemail = "Registro de Record Academico"
                    titulo = "Record Academico"
                    asunto = u"ADICION DE RECORD ACADEMICO"

                    datos = {'sistema': u'SGA - UNEMI',
                             'titulo': titulo,
                             'nombrepersona': personasesion.nombre_completo_inverso(),
                             'inscrito': record.inscripcion,
                             'record' : record,
                             'asunto': asunto, 'periodo': periodo}

                    send_html_mail(tituloemail, "emails/notificacion_addrecordacademico.html", datos,
                                   lista_email_cco,
                                   [], [], cuenta=CUENTAS_CORREOS[0][1])
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos. "})

        elif action == 'addrecordhomologada':
            try:
                inscripcion = Inscripcion.objects.get(pk=request.POST['id'])
                f = RecordAcademicoForm(request.POST)
                if f.is_valid():
                    if not inscripcion.recordacademico_set.filter(asignatura=f.cleaned_data['asignatura']).exists():
                        record = RecordAcademico(inscripcion=inscripcion,
                                                 asignatura=f.cleaned_data['asignatura'],
                                                 modulomalla=inscripcion.asignatura_en_modulomalla(f.cleaned_data['asignatura']),
                                                 # asignaturamalla=inscripcion.asignatura_en_asignaturamalla(f.cleaned_data['asignatura']),
                                                 nota=f.cleaned_data['nota'],
                                                 fecha=f.cleaned_data['fecha'],
                                                 aprobada=True,
                                                 noaplica=f.cleaned_data['noaplica'],
                                                 convalidacion=f.cleaned_data['convalidacion'],
                                                 pendiente=False,
                                                 valida=f.cleaned_data['valida'],
                                                 validapromedio=f.cleaned_data['validapromedio'],
                                                 creditos=f.cleaned_data['creditos'],
                                                 horas=f.cleaned_data['horas'],
                                                 homologada=f.cleaned_data['homologada'],
                                                 observaciones=f.cleaned_data['observaciones'])
                        record.save(request)
                    else:
                        recordexistente = inscripcion.recordacademico_set.filter(asignatura=f.cleaned_data['asignatura'])[0]
                        if recordexistente.historicorecordacademico_set.filter(fecha=f.cleaned_data['fecha']).exists():
                            return JsonResponse({"result": "bad", "mensaje": u"Registro existente en esa fecha"})
                        record = HistoricoRecordAcademico(recordacademico=recordexistente,
                                                          inscripcion=inscripcion,
                                                          asignatura=f.cleaned_data['asignatura'],
                                                          nota=f.cleaned_data['nota'],
                                                          fecha=f.cleaned_data['fecha'],
                                                          aprobada=True,
                                                          noaplica=f.cleaned_data['noaplica'],
                                                          convalidacion=f.cleaned_data['convalidacion'],
                                                          pendiente=False,
                                                          valida=f.cleaned_data['valida'],
                                                          validapromedio=f.cleaned_data['validapromedio'],
                                                          creditos=f.cleaned_data['creditos'],
                                                          horas=f.cleaned_data['horas'],
                                                          homologada=f.cleaned_data['homologada'],
                                                          observaciones=f.cleaned_data['observaciones'])
                        record.save(request)
                    record.actualizar()
                    inscripcion.actualizar_nivel()
                    inscripcion.actualiza_matriculas(f.cleaned_data['asignatura'])
                    log(u'Adiciono record academico: %s - %s' % (record, record.inscripcion.persona), request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos. "})

        elif action == 'addhistorico':
            try:
                record = RecordAcademico.objects.get(pk=request.POST['id'])
                f = HistoricoRecordAcademicoForm(request.POST)
                if f.is_valid():
                    if record.historicorecordacademico_set.filter(fecha=f.cleaned_data['fecha']).exists():
                        return JsonResponse({"result": "bad", "mensaje": u"Registro existente en esa fecha"})
                    historico = HistoricoRecordAcademico(recordacademico=record,
                                                         inscripcion=record.inscripcion,
                                                         asignatura=f.cleaned_data['asignatura'],
                                                         modulomalla=record.inscripcion.asignatura_en_modulomalla(f.cleaned_data['asignatura']),
                                                         asignaturamalla=record.inscripcion.asignatura_en_asignaturamalla(f.cleaned_data['asignatura']),
                                                         nota=f.cleaned_data['nota'],
                                                         asistencia=f.cleaned_data['asistencia'],
                                                         fecha=f.cleaned_data['fecha'],
                                                         aprobada=f.cleaned_data['aprobada'],
                                                         convalidacion=f.cleaned_data['convalidacion'],
                                                         noaplica=f.cleaned_data['noaplica'],
                                                         pendiente=False,
                                                         creditos=f.cleaned_data['creditos'],
                                                         horas=f.cleaned_data['horas'],
                                                         homologada=f.cleaned_data['homologada'],
                                                         valida=f.cleaned_data['valida'],
                                                         validapromedio=f.cleaned_data['validapromedio'],
                                                         observaciones=f.cleaned_data['observaciones'],
                                                         suficiencia=f.cleaned_data['suficiencia'])
                    historico.save(request)
                    historico.actualizar()
                    record.inscripcion.actualizar_nivel()
                    record.inscripcion.actualiza_matriculas(record.asignatura)
                    log(u'Adiciono historico y registro: %s - %s' % (historico, historico.inscripcion.persona), request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'edithistorico':
            try:
                puede_editar = variable_valor('ADICIONAR_RECORD')
                historico = HistoricoRecordAcademico.objects.get(pk=request.POST['id'])
                f = HistoricoRecordAcademicoForm(request.POST)
                if f.is_valid():
                    if HistoricoRecordAcademico.objects.filter(inscripcion=historico.inscripcion, asignatura=historico.asignatura, fecha=f.cleaned_data['fecha']).exclude(id=historico.id).exists():
                        return JsonResponse({"result": "bad", "mensaje": u"Registro existente en esa fecha."})
                    historico.fecha = f.cleaned_data['fecha']
                    if not historico.tiene_acta_nivel() or puede_editar:
                        historico.nota = f.cleaned_data['nota']
                        historico.asistencia = f.cleaned_data['asistencia']
                    historico.aprobada = f.cleaned_data['aprobada']
                    historico.noaplica = f.cleaned_data['noaplica']
                    historico.creditos = f.cleaned_data['creditos']
                    historico.horas = f.cleaned_data['horas']
                    historico.homologada = f.cleaned_data['homologada']
                    historico.convalidacion = f.cleaned_data['convalidacion']
                    historico.observaciones = f.cleaned_data['observaciones']
                    historico.valida = f.cleaned_data['valida']
                    historico.validapromedio = f.cleaned_data['validapromedio']
                    historico.suficiencia = f.cleaned_data['suficiencia']
                    historico.save(request)
                    historico.actualizar()
                    historico.inscripcion.actualizar_nivel()
                    historico.inscripcion.actualiza_matriculas(historico.asignatura)
                    log(u'Modifico historico de record academico: %s - %s' % (historico, historico.inscripcion.persona), request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'delrecord':
            try:
                record = RecordAcademico.objects.get(pk=request.POST['id'])
                asignatura = record.asignatura
                inscripcion = record.inscripcion
                homologacion = record.homologacioninscripcion_set.all()
                homologacion.delete()
                convalidacion = record.convalidacioninscripcion_set.all()
                convalidacion.delete()
                historico = record.historicorecordacademico_set.all()
                historico.delete()
                log(u'Elimino registro academico: %s - %s' % (record, record.inscripcion.persona), request, "del")
                record.delete()
                inscripcion.actualizar_nivel()
                inscripcion.actualiza_matriculas(asignatura)
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'delhistorico':
            try:
                historico = HistoricoRecordAcademico.objects.get(pk=request.POST['id'])
                asignatura = historico.asignatura
                inscripcion = historico.inscripcion
                if historico.convalidacion:
                    convalidacion = historico.convalidacioninscripcion_set.all()
                    convalidacion.delete()
                if historico.homologada:
                    homologacion = historico.homologacioninscripcion_set.all()
                    homologacion.delete()
                log(u'Elimino historico de registro academico: %s - %s' % (historico, historico.inscripcion.persona), request, "del")
                historico.delete()
                if HistoricoRecordAcademico.objects.filter(asignatura=asignatura, inscripcion=inscripcion).exists():
                    historico = HistoricoRecordAcademico.objects.filter(asignatura=asignatura, inscripcion=inscripcion)[0]
                    historico.actualizar()
                else:
                    record = RecordAcademico.objects.filter(asignatura=asignatura, inscripcion=inscripcion)
                    record.delete()
                inscripcion.actualizar_nivel()
                inscripcion.actualiza_matriculas(asignatura)
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'convalidar':
            try:
                convalidacion = ConvalidacionInscripcion.objects.get(pk=request.POST['id'])
                f = ConvalidacionInscripcionForm(request.POST, request.FILES)
                if f.is_valid():
                    convalidacion.centro = f.cleaned_data['centro']
                    convalidacion.carrera = f.cleaned_data['carrera']
                    convalidacion.asignatura = f.cleaned_data['asignatura']
                    convalidacion.anno = f.cleaned_data['anno']
                    convalidacion.nota_ant = f.cleaned_data['nota_ant']
                    convalidacion.nota_act = f.cleaned_data['nota_act']
                    convalidacion.observaciones = f.cleaned_data['observaciones']
                    convalidacion.creditos = f.cleaned_data['creditos']
                    convalidacion.save(request)
                    if 'archivo' in request.FILES:
                        archivo = request.FILES['archivo']
                        archivo._name = generar_nombre("archivo_", archivo._name)
                        convalidacion.archivo = archivo
                        convalidacion.save(request)
                    log(u'Adiciono convalidacion: %s - %s' % (convalidacion, convalidacion.record.inscripcion.persona), request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'homologar':
            try:
                homologacion = HomologacionInscripcion.objects.get(pk=request.POST['id'])
                f = HomologacionInscripcionForm(request.POST, request.FILES)
                if f.is_valid():
                    homologacion.carrera = f.cleaned_data['carrera']
                    homologacion.modalidad = f.cleaned_data['modalidad']
                    homologacion.asignatura = f.cleaned_data['asignatura']
                    homologacion.fecha = f.cleaned_data['fecha']
                    homologacion.nota_ant = f.cleaned_data['nota_ant']
                    homologacion.observaciones = f.cleaned_data['observaciones']
                    homologacion.creditos = f.cleaned_data['creditos']
                    homologacion.save(request)
                    if 'archivo' in request.FILES:
                        archivo = request.FILES['archivo']
                        archivo._name = generar_nombre("archivo_", archivo._name)
                        homologacion.archivo = archivo
                        homologacion.save(request)
                    log(u'Adiciono homologacion: %s - %s' % (homologacion, homologacion.record.inscripcion.persona), request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'cargarfoto':
            try:
                inscripcion = Inscripcion.objects.get(pk=request.POST['id'])
                form = CargarFotoForm(inscripcion, request.FILES)
                if form.is_valid():
                    persona = inscripcion.persona
                    newfile = request.FILES['foto']
                    newfile._name = generar_nombre("foto_", newfile._name)
                    foto = persona.foto()
                    if foto:
                        foto.foto = newfile
                    else:
                        foto = FotoPersona(persona=persona,
                                           foto=newfile)
                    foto.save(request)
                    make_thumb_picture(persona)
                    if GENERAR_TUMBAIL:
                        make_thumb_fotopersona(persona)
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos. "})

        elif action == 'cambiogrupo':
            try:
                inscripcion = Inscripcion.objects.get(pk=request.POST['id'])
                if inscripcion.matriculado():
                    return JsonResponse({"result": "bad", "mensaje": u"Existe una matricula activa."})
                f = CambioGrupoForm(request.POST)
                if f.is_valid():
                    inscripcion.inscripcion_grupo(f.cleaned_data['grupo'])
                    log(u'Cambio de grupo: %s - %s' % (inscripcion.persona, f.cleaned_data['grupo']), request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'add':
            try:
                f = InscripcionForm(request.POST, request.FILES)
                isPPL = False
                archivoppl = None
                if 'ppl' in request.POST:
                    isPPL = True
                    f.fields['fechaingresoppl'].required=True
                    if 'archivoppl' in request.FILES:
                        arch = request.FILES['archivoppl']
                        extension = arch._name.split('.')
                        tam = len(extension)
                        exte = extension[tam - 1]
                        if arch.size > 4194304:
                            raise NameError(u"Tamaño del archivo es mayor a 4 Mb")
                        if not exte.lower() == 'pdf':
                            raise NameError(u"Solo se permiten archivos .pdf")
                        archivoppl = request.FILES['archivoppl']
                        archivoppl._name = generar_nombre("archivoppl_", archivoppl._name)
                if not f.is_valid():
                    raise NameError('Formulario incorrecto')
                fechnacimiento = str(f.cleaned_data['nacimiento']).split('-')
                fechnacimiento = int(fechnacimiento[0])
                anioactual = datetime.now()
                ahora = int(anioactual.year)
                restafecha = ahora - fechnacimiento
                if restafecha < 16:
                    return JsonResponse({"result": "bad", "mensaje": u"La fecha de nacimiento no es valida."})
                if f.cleaned_data['cedula'] and Persona.objects.filter(cedula=f.cleaned_data['cedula']).exists():
                    if Inscripcion.objects.filter(persona__cedula=f.cleaned_data['cedula'], carrera=f.cleaned_data['carrera']).exists():
                        return JsonResponse({"result": "bad", "mensaje": u"El numero de cedula o pasaporte ya esta registrado en la carrera seleccionada."})
                if f.cleaned_data['pasaporte'] and Persona.objects.filter(pasaporte=f.cleaned_data['pasaporte']).exists():
                    if Inscripcion.objects.filter(persona__cedula=f.cleaned_data['pasaporte'], carrera=f.cleaned_data['carrera']).exists():
                        return JsonResponse({"result": "bad", "mensaje": u"El numero de pasaporte ya esta registrado en la carrera seleccionada."})
                if not f.cleaned_data['cedula'] and not f.cleaned_data['pasaporte']:
                    return JsonResponse({"result": "bad", "mensaje": u"Debe especificar un numero de identificación."})
                if f.cleaned_data['cedula'] == "" and f.cleaned_data['pasaporte'] == "":
                    return JsonResponse({"result": "bad", "mensaje": u"Debe especificar un numero de identificación."})
                persona=None
                if f.cleaned_data['cedula'] != "" and f.cleaned_data['pasaporte'] != "":
                    if Persona.objects.filter(Q(cedula=f.cleaned_data['cedula'])|Q(pasaporte=f.cleaned_data['pasaporte'])).exists():
                        persona = Persona.objects.filter(Q(cedula=f.cleaned_data['cedula'])|Q(pasaporte=f.cleaned_data['pasaporte']))
                elif f.cleaned_data['cedula'] != "":
                    if Persona.objects.filter(Q(cedula=f.cleaned_data['cedula'])|Q(pasaporte=f.cleaned_data['cedula'])).exists():
                        persona = Persona.objects.filter(Q(cedula=f.cleaned_data['cedula'])|Q(pasaporte=f.cleaned_data['cedula']))
                elif f.cleaned_data['pasaporte'] != "":
                    if Persona.objects.filter(Q(cedula=f.cleaned_data['pasaporte'])|Q(pasaporte=f.cleaned_data['pasaporte'])).exists():
                        persona = Persona.objects.filter(Q(cedula=f.cleaned_data['pasaporte'])|Q(pasaporte=f.cleaned_data['pasaporte']))

                if persona:
                    if persona.count() > 1:
                        return JsonResponse({"result": "bad", "mensaje": u"Existe màs de una persona con esa identificación."})
                    persona = persona[0]
                    persona.nombres = f.cleaned_data['nombres']
                    persona.apellido1 = f.cleaned_data['apellido1']
                    persona.apellido2 = f.cleaned_data['apellido2']
                    persona.cedula = f.cleaned_data['cedula']
                    persona.pasaporte = f.cleaned_data['pasaporte']
                    persona.nacimiento = f.cleaned_data['nacimiento']
                    persona.sexo = f.cleaned_data['sexo']
                    persona.paisnacimiento = f.cleaned_data['paisnacimiento']
                    persona.provincianacimiento = f.cleaned_data['provincianacimiento']
                    persona.cantonnacimiento = f.cleaned_data['cantonnacimiento']
                    persona.parroquianacimiento = f.cleaned_data['parroquianacimiento']
                    persona.nacionalidad = f.cleaned_data['nacionalidad']
                    persona.pais = f.cleaned_data['pais']
                    persona.provincia = f.cleaned_data['provincia']
                    persona.canton = f.cleaned_data['canton']
                    persona.parroquia = f.cleaned_data['parroquia']
                    persona.sector = f.cleaned_data['sector']
                    persona.direccion = f.cleaned_data['direccion']
                    persona.direccion2 = f.cleaned_data['direccion2']
                    persona.num_direccion = f.cleaned_data['num_direccion']
                    persona.telefono = f.cleaned_data['telefono']
                    persona.telefono_conv = f.cleaned_data['telefono_conv']
                    persona.email = f.cleaned_data['email']
                    persona.sangre = f.cleaned_data['sangre']
                    persona.lgtbi = f.cleaned_data['lgtbi']
                    # persona.ppl = f.cleaned_data['ppl']
                    # persona.observacionppl = f.cleaned_data['observacionppl'] if f.cleaned_data['observacionppl'] else None
                    persona.save(request)
                    if not persona.usuario:
                        username = calculate_username(persona)
                        generar_usuario(persona, username, ALUMNOS_GROUP_ID)
                    else:
                        username = persona.usuario.__str__()
                else:
                    persona = Persona(nombres=f.cleaned_data['nombres'],
                                      apellido1=f.cleaned_data['apellido1'],
                                      apellido2=f.cleaned_data['apellido2'],
                                      cedula=f.cleaned_data['cedula'],
                                      pasaporte=f.cleaned_data['pasaporte'],
                                      nacimiento=f.cleaned_data['nacimiento'],
                                      sexo=f.cleaned_data['sexo'],
                                      paisnacimiento=f.cleaned_data['paisnacimiento'],
                                      provincianacimiento=f.cleaned_data['provincianacimiento'],
                                      cantonnacimiento=f.cleaned_data['cantonnacimiento'],
                                      parroquianacimiento=f.cleaned_data['parroquianacimiento'],
                                      nacionalidad=f.cleaned_data['nacionalidad'],
                                      pais=f.cleaned_data['pais'],
                                      provincia=f.cleaned_data['provincia'],
                                      canton=f.cleaned_data['canton'],
                                      parroquia=f.cleaned_data['parroquia'],
                                      sector=f.cleaned_data['sector'],
                                      direccion=f.cleaned_data['direccion'],
                                      direccion2=f.cleaned_data['direccion2'],
                                      num_direccion=f.cleaned_data['num_direccion'],
                                      telefono=f.cleaned_data['telefono'],
                                      telefono_conv=f.cleaned_data['telefono_conv'],
                                      email=f.cleaned_data['email'],
                                      sangre=f.cleaned_data['sangre'],
                                      lgtbi=f.cleaned_data['lgtbi'],
                                      ppl=f.cleaned_data['ppl'],
                                      observacionppl=f.cleaned_data['observacionppl'] )
                    persona.save(request)
                    username = calculate_username(persona)
                    generar_usuario(persona, username, ALUMNOS_GROUP_ID)

                if EMAIL_INSTITUCIONAL_AUTOMATICO:
                    persona.emailinst = username + '@' + EMAIL_DOMAIN
                    persona.save(request)
                if UTILIZA_GRUPOS_ALUMNOS:
                    grupo = f.cleaned_data['grupo']
                    carrera = grupo.carrera
                    sesion = grupo.sesion
                    modalidad = grupo.modalidad
                    sede = grupo.sede
                else:
                    carrera = f.cleaned_data['carrera']
                    sesion = f.cleaned_data['sesion']
                    modalidad = f.cleaned_data['modalidad']
                    sede = f.cleaned_data['sede']
                if not persona.usuario.is_superuser:
                    if persona.inscripcion_set.values('id').exclude(carrera__coordinacion__in=[9,7]).count() > 2:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": "bad", "mensaje": u"Llego al limite de inscripcion en las carreras."})
                inscripcion = Inscripcion(persona=persona,
                                          fecha=f.cleaned_data['fecha'],
                                          # colegio=f.cleaned_data['colegio'],
                                          especialidad=f.cleaned_data['especialidad'],
                                          identificador=f.cleaned_data['identificador'],
                                          centroinformacion=f.cleaned_data['centroinformacion'],
                                          carrera=carrera,
                                          modalidad=modalidad,
                                          sesion=sesion,
                                          sede=sede,
                                          estado_gratuidad=1,
                                          porcentaje_perdida_gratuidad=0)
                if f.cleaned_data['unidadeducativa']:
                    inscripcion.unidadeducativa_id = f.cleaned_data['unidadeducativa']
                inscripcion.save(request)
                if isPPL:
                    if HistorialPersonaPPL.objects.filter(fechaingreso=f.cleaned_data['fechaingresoppl'], persona=persona, inscripcion=inscripcion).exists():
                        historialppl = HistorialPersonaPPL.objects.filter(fechaingreso=f.cleaned_data['fechaingresoppl'], persona=persona, inscripcion=inscripcion).first()
                        historialppl.observacion = f.cleaned_data['observacionppl'] if f.cleaned_data['observacionppl'] else historialppl.observacion
                        historialppl.archivo = archivoppl if archivoppl else historialppl.archivo
                        historialppl.centrorehabilitacion = f.cleaned_data['centrorehabilitacion'] if f.cleaned_data['centrorehabilitacion'] else historialppl.centrorehabilitacion
                        historialppl.lidereducativo = f.cleaned_data['lidereducativo'] if f.cleaned_data['lidereducativo'] else historialppl.lidereducativo
                        historialppl.correolidereducativo = f.cleaned_data['correolidereducativo'] if f.cleaned_data['correolidereducativo'] else historialppl.correolidereducativo
                        historialppl.telefonolidereducativo = f.cleaned_data['telefonolidereducativo'] if f.cleaned_data['telefonolidereducativo'] else historialppl.telefonolidereducativo
                    else:
                        historialppl = HistorialPersonaPPL(persona=persona,
                                                           inscripcion=inscripcion,
                                                           observacion=f.cleaned_data['observacionppl'] if f.cleaned_data['observacionppl'] else None,
                                                           archivo=archivoppl,
                                                           fechaingreso=f.cleaned_data['fechaingresoppl'],
                                                           centrorehabilitacion=f.cleaned_data['centrorehabilitacion'] if f.cleaned_data['centrorehabilitacion'] else None,
                                                           lidereducativo=f.cleaned_data['lidereducativo'] if f.cleaned_data['lidereducativo'] else None,
                                                           correolidereducativo=f.cleaned_data['correolidereducativo'] if f.cleaned_data['correolidereducativo'] else None,
                                                           telefonolidereducativo=f.cleaned_data['telefonolidereducativo'] if f.cleaned_data['telefonolidereducativo'] else None,
                                                           )
                    historialppl.save(request)
                if not inscripcion.coordinacion:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"No existe una coordinacion para esta carrera en esta sede."})
                # DOCUMENTOS DE INSCRIPCION
                documentos = DocumentosDeInscripcion(inscripcion=inscripcion,
                                                     titulo=f.cleaned_data['titulo'],
                                                     acta=f.cleaned_data['acta'],
                                                     cedula=f.cleaned_data['cedula2'],
                                                     votacion=f.cleaned_data['votacion'],
                                                     actaconv=f.cleaned_data['actaconv'],
                                                     partida_nac=f.cleaned_data['partida_nac'],
                                                     pre=f.cleaned_data['prenivelacion'],
                                                     observaciones_pre=f.cleaned_data['observacionespre'],
                                                     fotos=f.cleaned_data['fotos'])
                documentos.save(request)
                # DOCUMENTOS CONDUCCION
                inscripciontesdrive = InscripcionTesDrive(inscripcion=inscripcion,
                                                          licencia=f.cleaned_data['licencia'],
                                                          record=f.cleaned_data['record'],
                                                          certificado_tipo_sangre=f.cleaned_data['certificado_tipo_sangre'],
                                                          prueba_psicosensometrica=f.cleaned_data['prueba_psicosensometrica'],
                                                          certificado_estudios=f.cleaned_data['certificado_estudios'])
                inscripciontesdrive.save(request)
                # SEGUIMIENTO LABORAL
                if f.cleaned_data['trabaja']:
                    trabajo = SeguimientoEstudiante(persona=persona,
                                                    empresa=f.cleaned_data['empresa'],
                                                    ocupacion=f.cleaned_data['ocupacion'],
                                                    responsabilidades='',
                                                    personascargo=0,
                                                    ejerce=False,
                                                    fecha=f.cleaned_data['fecha_ingreso'],
                                                    telefono=f.cleaned_data['telefono_trabajo'])
                    trabajo.save(request)
                # REGISTRO TIPO DE INSCRIPCION
                if USA_TIPOS_INSCRIPCIONES:
                    inscripciontipoinscripcion = InscripcionTipoInscripcion(inscripcion=inscripcion,
                                                                            tipoinscripcion_id=TIPO_INSCRIPCION_INICIAL)
                    inscripciontipoinscripcion.save(request)
                # PREGUNTAS
                preguntasinscripcion = inscripcion.preguntas_inscripcion()
                if f.cleaned_data['comoseinformo']:
                    preguntasinscripcion.comoseinformo = f.cleaned_data['comoseinformo']
                if f.cleaned_data['razonesmotivaron']:
                    preguntasinscripcion.razonesmotivaron = f.cleaned_data['razonesmotivaron']
                if f.cleaned_data['comoseinformootras']:
                    preguntasinscripcion.comoseinformootras = f.cleaned_data['comoseinformootras']
                preguntasinscripcion.save(request)
                # PREINSCRITO
                if 'preinscrito_id' in request.POST:
                    preinscrito = PreInscrito.objects.get(pk=request.POST['preinscrito_id'])
                    preinscrito.inscripcion = inscripcion
                    preinscrito.save(request)
                # PERFIL DE USUARIO
                persona.crear_perfil(inscripcion=inscripcion)
                perfil=persona.mi_perfil()
                perfil.raza=(f.cleaned_data['raza'])
                perfil.nacionalidadindigena = (f.cleaned_data['nacionalidadindigena'])
                perfil.save(request)
                if UTILIZA_GRUPOS_ALUMNOS:
                    inscripcion.inscripcion_grupo(f.cleaned_data['grupo'])
                inscripcion.persona.mi_perfil()
                inscripcion.malla_inscripcion()
                inscripcion.actualizar_nivel()
                fichasocioeconomica = ficha_socioeconomica(inscripcion.persona)
                log(u'Adiciono inscripcion: %s' % inscripcion, request, "add")

                lista_email_cco = ['sga@unemi.edu.ec']
                tituloemail = "Registro de Inscripcion"
                titulo = "Inscripcion"
                asunto = u"ADICION DE INSCRIPCION"

                datos = {'sistema': u'SGA - UNEMI',
                         'titulo': titulo,
                         'nombrepersona': personasesion.nombre_completo_inverso(),
                         'inscrito': inscripcion,
                         'asunto': asunto, 'periodo': periodo}


                send_html_mail(tituloemail, "emails/notificacion_addinscripcion.html", datos,
                               lista_email_cco,
                               [], [],cuenta=CUENTAS_CORREOS[0][1])

                #send_html_mail("Crear nueva cuenta de correo ", "emails/nuevacuentacorreo.html", {'sistema': request.session['nombresistema'], 'persona': persona, 't': miinstitucion(), 'tipo_usuario': 'ESTUDIANTE', 'inscripcion': inscripcion}, lista_correo([variable_valor('ADMINISTRADOR_CORREO_GROUP_ID')]), [], cuenta=CUENTAS_CORREOS[4][1])
                persona.creacion_persona(request.session['nombresistema'], personasesion)
                messages.success(request, 'Guardado con exito')
                return JsonResponse({"result": "ok", "id": inscripcion.id})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos. %s" % ex.__str__()})

        elif action == 'adicionarotracarrera':
            try:
                f = NuevaInscripcionForm(request.POST)
                if f.is_valid() and not Inscripcion.objects.filter(id=int(request.POST['id']), carrera=f.cleaned_data['carrera']).exists():
                    inscripcion = Inscripcion.objects.get(pk=int(request.POST['id']))
                    carrera = f.cleaned_data['carrera']
                    sesion = f.cleaned_data['sesion']
                    modalidad = f.cleaned_data['modalidad']
                    sede = f.cleaned_data['sede']
                    if Inscripcion.objects.filter(persona=inscripcion.persona, status=True).exclude(coordinacion__id__in=[7,9]).count() > 5:
                        return JsonResponse({"result": "bad", "mensaje": u"Ya se encuentra registrado en mas de 1 carrera."})
                    if Inscripcion.objects.filter(persona=inscripcion.persona, carrera=carrera).exists():
                        return JsonResponse({"result": "bad", "mensaje": u"Ya se encuentra registrado en esa carrera."})
                    nuevainscripcion = Inscripcion(persona=inscripcion.persona,
                                                   fecha=f.cleaned_data['fecha'],
                                                   colegio=inscripcion.colegio,
                                                   especialidad=inscripcion.especialidad,
                                                   estado_gratuidad=inscripcion.estado_gratuidad,
                                                   porcentaje_perdida_gratuidad=inscripcion.porcentaje_perdida_gratuidad,
                                                   carrera=carrera,
                                                   modalidad=modalidad,
                                                   sesion=sesion,
                                                   sede=sede)
                    nuevainscripcion.save(request)
                    if f.cleaned_data['copiarecord']:
                        for record in inscripcion.recordacademico():
                            nuevorecord = RecordAcademico(inscripcion=nuevainscripcion,
                                                          asignatura=record.asignatura,
                                                          nota=record.nota,
                                                          asistencia=record.asistencia,
                                                          fecha=record.fecha,
                                                          noaplica=record.noaplica,
                                                          convalidacion=False,
                                                          homologada=True,
                                                          aprobada=record.aprobada,
                                                          pendiente=record.pendiente,
                                                          creditos=record.creditos,
                                                          horas=record.horas,
                                                          valida=record.valida,
                                                          observaciones=record.observaciones)
                            nuevorecord.save(request)
                            for hrecord in record.historicorecordacademico_set.all():
                                newrecordhisto = HistoricoRecordAcademico(recordacademico=nuevorecord,
                                                                          inscripcion=nuevainscripcion,
                                                                          modulomalla=nuevorecord.modulomalla,
                                                                          asignaturamalla=nuevorecord.asignaturamalla,
                                                                          asignatura=hrecord.asignatura,
                                                                          nota=hrecord.nota,
                                                                          asistencia=hrecord.asistencia,
                                                                          sinasistencia=hrecord.sinasistencia,
                                                                          fecha=hrecord.fecha,
                                                                          noaplica=hrecord.noaplica,
                                                                          aprobada=hrecord.aprobada,
                                                                          convalidacion=hrecord.convalidacion,
                                                                          homologada=hrecord.homologada,
                                                                          pendiente=hrecord.pendiente,
                                                                          creditos=hrecord.creditos,
                                                                          horas=hrecord.horas,
                                                                          valida=hrecord.valida,
                                                                          validapromedio=hrecord.validapromedio,
                                                                          materiaregular=hrecord.materiaregular,
                                                                          materiacurso=hrecord.materiacurso,
                                                                          observaciones=hrecord.observaciones,
                                                                          completonota=hrecord.completonota,
                                                                          completoasistencia=hrecord.completoasistencia,
                                                                          suficiencia=hrecord.suficiencia)
                                newrecordhisto.save()
                            nuevorecord.actualizar()
                    if inscripcion.tipo_inscripcion():
                        tipoinscripcion = inscripcion.tipo_inscripcion().tipoinscripcion
                        nuevotipo = InscripcionTipoInscripcion(inscripcion=nuevainscripcion,
                                                               tipoinscripcion=tipoinscripcion)
                        nuevotipo.save(request)
                    hoy = datetime.now().date()
                    inscripciontesdrive = InscripcionTesDrive(inscripcion=nuevainscripcion,
                                                              licencia=f.cleaned_data['licencia'],
                                                              record=f.cleaned_data['record'],
                                                              certificado_tipo_sangre=f.cleaned_data['certificado_tipo_sangre'],
                                                              prueba_psicosensometrica=f.cleaned_data['prueba_psicosensometrica'],
                                                              certificado_estudios=f.cleaned_data['certificado_estudios'])
                    inscripciontesdrive.save(request)
                    nuevainscripcion.malla_inscripcion()
                    log(u'Adiciono inscripcion desde otra carrera: %s' % nuevainscripcion, request, "add")
                    documentos = DocumentosDeInscripcion(inscripcion=nuevainscripcion,
                                                         titulo=f.cleaned_data['titulo'],
                                                         acta=f.cleaned_data['acta'],
                                                         cedula=f.cleaned_data['cedula2'],
                                                         votacion=f.cleaned_data['votacion'],
                                                         actaconv=f.cleaned_data['actaconv'],
                                                         partida_nac=f.cleaned_data['partida_nac'],
                                                         pre=f.cleaned_data['prenivelacion'],
                                                         observaciones_pre=f.cleaned_data['observacionespre'],
                                                         fotos=f.cleaned_data['fotos'])
                    documentos.save(request)
                    inscripcion.persona.crear_perfil(inscripcion=nuevainscripcion)
                    fichasocioeconomica = ficha_socioeconomica(inscripcion.persona)
                    if inscripcion.tiene_perdida_gratuidad():
                        perdidas = PerdidaGratuidad.objects.filter(inscripcion=inscripcion,status=True)
                        for perdida in perdidas:
                            perdidagratuidad = PerdidaGratuidad(
                                inscripcion=nuevainscripcion,
                                motivo=perdida.motivo,
                                titulo=perdida.titulo,
                                observacion=perdida.observacion,
                                titulo_sniese=perdida.titulo_sniese,
                                cupo_aceptado_senescyt=perdida.cupo_aceptado_senescyt,
                                segunda_carrera_raes=perdida.segunda_carrera_raes,
                            )
                            perdidagratuidad.save(request)
                    return JsonResponse({"result": "ok", "id": nuevainscripcion.id})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'edit':
            try:
                f = InscripcionForm(request.POST)
                if f.is_valid():
                    inscripcion = Inscripcion.objects.get(pk=request.POST['id'])
                    persona = inscripcion.persona
                    perfil = persona.mi_perfil()
                    persona.nombres = f.cleaned_data['nombres']
                    persona.apellido1 = f.cleaned_data['apellido1']
                    persona.apellido2 = f.cleaned_data['apellido2']
                    persona.nacimiento = f.cleaned_data['nacimiento']
                    persona.paisnacimiento = f.cleaned_data['paisnacimiento']
                    persona.provincianacimiento = f.cleaned_data['provincianacimiento']
                    persona.cantonnacimiento = f.cleaned_data['cantonnacimiento']
                    persona.parroquianacimiento = f.cleaned_data['parroquianacimiento']
                    persona.nacionalidad = f.cleaned_data['nacionalidad']
                    persona.pais = f.cleaned_data['pais']
                    persona.provincia = f.cleaned_data['provincia']
                    persona.canton = f.cleaned_data['canton']
                    persona.parroquia = f.cleaned_data['parroquia']
                    persona.sexo = f.cleaned_data['sexo']
                    persona.provincia = f.cleaned_data['provincia']
                    persona.canton = f.cleaned_data['canton']
                    persona.parroquia = f.cleaned_data['parroquia']
                    persona.sector = f.cleaned_data['sector']
                    persona.direccion = f.cleaned_data['direccion']
                    persona.direccion2 = f.cleaned_data['direccion2']
                    persona.num_direccion = f.cleaned_data['num_direccion']
                    persona.telefono = f.cleaned_data['telefono']
                    persona.telefono_conv = f.cleaned_data['telefono_conv']
                    persona.email = f.cleaned_data['email']
                    persona.emailinst = f.cleaned_data['emailinst']
                    persona.sangre = f.cleaned_data['sangre']
                    persona.lgtbi = f.cleaned_data['lgtbi']
                    # persona.ppl = f.cleaned_data['ppl']
                    # persona.observacionppl = f.cleaned_data['observacionppl']
                    persona.save(request)
                    #DATOS UBE
                    perfil.raza = f.cleaned_data['raza']
                    perfil.nacionalidadindigena = f.cleaned_data['nacionalidadindigena']
                    perfil.save(request)
                    # DATOS DE LA INSCRIPCION
                    inscripcion.fecha = f.cleaned_data['fecha']
                    if not inscripcion.matriculado() and not inscripcion.graduado() and not inscripcion.egresado():
                        if UTILIZA_GRUPOS_ALUMNOS:
                            grupo = f.cleaned_data['grupo']
                            inscripcion.inscripcion_grupo(grupo)
                        else:
                            inscripcion.modalidad = f.cleaned_data['modalidad']
                            inscripcion.sede = f.cleaned_data['sede']
                    inscripcion.centroinformacion = f.cleaned_data['centroinformacion']
                    inscripcion.sesion = f.cleaned_data['sesion']
                    # inscripcion.colegio = f.cleaned_data['colegio']
                    # inscripcion.colegio = ''
                    if f.cleaned_data['unidadeducativa']:
                        if int(f.cleaned_data['unidadeducativa']) > 0:
                            inscripcion.unidadeducativa_id=f.cleaned_data['unidadeducativa']
                    inscripcion.especialidad_id = f.cleaned_data['especialidad']
                    inscripcion.identificador = f.cleaned_data['identificador']
                    if puede_realizar_accion_afirmativo(request, 'sga.puede_modificar_itinerario'):
                        inscripcion.itinerario = f.cleaned_data['itinerario']
                    inscripcion.save(request)
                    # ACTUALIZA LAS PREGUNTAS DE INSCRIPCION
                    preguntas = inscripcion.preguntas_inscripcion()
                    preguntas.comoseinformo = f.cleaned_data['comoseinformo']
                    preguntas.razonesmotivaron = f.cleaned_data['razonesmotivaron']
                    preguntas.comoseinformootras = f.cleaned_data['comoseinformootras']
                    preguntas.save(request)
                    # DOCUMENTOS
                    documentos = inscripcion.documentos_entregados()
                    documentos.acta = f.cleaned_data['acta']
                    documentos.cedula = f.cleaned_data['cedula2']
                    documentos.fotos = f.cleaned_data['fotos']
                    documentos.titulo = f.cleaned_data['titulo']
                    documentos.votacion = f.cleaned_data['votacion']
                    documentos.actaconv = f.cleaned_data['actaconv']
                    documentos.partida_nac = f.cleaned_data['partida_nac']
                    # PRENIVELACION
                    documentos.pre = f.cleaned_data['prenivelacion']
                    documentos.observaciones_pre = f.cleaned_data['observacionespre']
                    documentos.save(request)
                    # OTROS DOCUMENTOS
                    tesdrive = inscripcion.documentos_tesdrive()
                    tesdrive.licencia = f.cleaned_data['licencia']
                    tesdrive.record = f.cleaned_data['record']
                    tesdrive.certificado_tipo_sangre = f.cleaned_data['certificado_tipo_sangre']
                    tesdrive.prueba_psicosensometrica = f.cleaned_data['prueba_psicosensometrica']
                    tesdrive.certificado_estudios = f.cleaned_data['certificado_estudios']
                    tesdrive.save(request)
                    # OTRAS ACCIONES
                    inscripcion.actualizar_nivel()
                    messages.success(request, 'Guardado con exito')
                    log(u'Modifico de inscripcion: %s' % inscripcion, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "mensaje": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'novalidar':
            try:
                record = RecordAcademico.objects.get(pk=request.POST['id'])
                f = ConsiderarForm(request.POST)
                if f.is_valid():
                    motivo = f.cleaned_data['motivo']
                    record.valida = False
                    record.save(request)
                    historico = record.mi_historico()
                    if historico:
                        historico.valida = record.valida
                        historico.save(request)
                    log(u'No considerar creditos: %s - %s - %s' % (record.inscripcion, record, motivo), request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'validar':
            try:
                record = RecordAcademico.objects.get(pk=request.POST['id'])
                f = ConsiderarForm(request.POST)
                if f.is_valid():
                    motivo = f.cleaned_data['motivo']
                    record.valida = True
                    record.save(request)
                    historico = record.mi_historico()
                    if historico:
                        historico.valida = record.valida
                        historico.save(request)
                    log(u'Considerar creditos: %s - %s - %s' % (record.inscripcion, record, motivo), request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'novalidarpromedio':
            try:
                record = RecordAcademico.objects.get(pk=request.POST['id'])
                f = ConsiderarForm(request.POST)
                if f.is_valid():
                    motivo = f.cleaned_data['motivo']
                    record.validapromedio = False
                    record.save(request)
                    historico = record.mi_historico()
                    if historico:
                        historico.validapromedio = record.validapromedio
                        historico.save(request)

                    log(u'No considerar promedio: %s - %s - %s' % (record.inscripcion, record, motivo), request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'validarpromedio':
            try:
                record = RecordAcademico.objects.get(pk=request.POST['id'])
                f = ConsiderarForm(request.POST)
                if f.is_valid():
                    motivo = f.cleaned_data['motivo']
                    record.validapromedio = True
                    record.save(request)
                    historico = record.mi_historico()
                    if historico:
                        historico.validapromedio = record.validapromedio
                        historico.save(request)
                    log(u'Considerar promedio: %s - %s - %s' % (record.inscripcion, record, motivo), request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'adddocumento':
            try:
                form = DocumentoInscripcionForm(request.POST, request.FILES)
                inscripcion = Inscripcion.objects.get(pk=request.POST['inscripcion'])
                if form.is_valid():
                    nfile = request.FILES['archivo']
                    nfile._name = generar_nombre("documentos_", nfile._name)
                    archivo = Archivo(nombre=form.cleaned_data['nombre'],
                                      fecha=datetime.now().date(),
                                      archivo=nfile,
                                      tipo_id=int(request.POST['tipo']),#ARCHIVO_TIPO_GENERAL,
                                      inscripcion=inscripcion)
                    archivo.save(request)
                    log(u'Adiciono documento de inscripcion: %s - %s' % (inscripcion, archivo), request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'importar':
            try:
                form = ImportarArchivoXLSForm(request.POST, request.FILES)
                if form.is_valid():
                    hoy = datetime.now().date()
                    nfile = request.FILES['archivo']
                    nfile._name = generar_nombre("importacion_", nfile._name)
                    archivo = Archivo(nombre='IMPORTACION INSCRIPCIONES',
                                      fecha=datetime.now().date(),
                                      archivo=nfile,
                                      tipo_id=ARCHIVO_TIPO_GENERAL)
                    archivo.save(request)
                    workbook = xlrd.open_workbook(archivo.archivo.file.name)
                    sheet = workbook.sheet_by_index(0)
                    linea = 1
                    for rowx in range(sheet.nrows):
                        puntosalva = transaction.savepoint()
                        try:
                            cols = sheet.row_values(rowx)
                            cedula = cols[0].strip().upper()
                            bandera = True
                            if Persona.objects.filter(cedula=cols[0]).exists():
                                bandera = False
                            if bandera:
                                if cedula and Persona.objects.filter(cedula=cedula).exists():
                                    return JsonResponse({"result": "bad", "mensaje": u"Existe una persona con esta identificación: %s." % cedula})
                                persona = Persona(cedula=cedula,
                                                  apellido1=cols[1],
                                                  apellido2=cols[2],
                                                  nombres=cols[3],
                                                  sexo_id=int(cols[4]),
                                                  nacimiento=xlrd.xldate.xldate_as_datetime(cols[5], workbook.datemode).date(),
                                                  provincianac=cols[6],
                                                  cantonnac=cols[7],
                                                  ciudad=cols[8],
                                                  email=cols[9],
                                                  telefono_conv=cols[10],
                                                  telefono=cols[11])
                                persona.save(request)
                                username = calculate_username(persona)
                                usuario = generar_usuario(persona, username, ALUMNOS_GROUP_ID)
                                if EMAIL_INSTITUCIONAL_AUTOMATICO:
                                    persona.emailinst = username + '@' + EMAIL_DOMAIN
                                    persona.save(request)
                                grupo = None
                                if UTILIZA_GRUPOS_ALUMNOS:
                                    grupo = Grupo.objects.get(pk=int(cols[15]))
                                    carrera = grupo.carrera
                                    sesion = grupo.sesion
                                    modalidad = grupo.modalidad
                                    sede = grupo.sede
                                else:
                                    sesion = Sesion.objects.get(pk=int(cols[12]))
                                    carrera = Carrera.objects.get(pk=int(cols[13]))
                                    modalidad = Modalidad.objects.get(pk=int(cols[14]))
                                    sede = Sede.objects.get(pk=int(cols[15]))
                            else:
                                persona = Persona.objects.filter(cedula=cols[0])[0]
                                sesion = Sesion.objects.get(pk=int(cols[12]))
                                carrera = Carrera.objects.get(pk=int(cols[13]))
                                modalidad = Modalidad.objects.get(pk=int(cols[14]))
                                sede = Sede.objects.get(pk=int(cols[15]))
                            inscripcion = Inscripcion(persona=persona,
                                                      fecha=xlrd.xldate.xldate_as_datetime(cols[18], workbook.datemode).date(),
                                                      carrera=carrera,
                                                      modalidad=modalidad,
                                                      sesion=sesion,
                                                      sede=sede,
                                                      colegio=cols[16])
                            inscripcion.save(request)
                            persona.crear_perfil(inscripcion=inscripcion)
                            documentos = DocumentosDeInscripcion(inscripcion=inscripcion,
                                                                 titulo=False,
                                                                 acta=False,
                                                                 cedula=False,
                                                                 votacion=False,
                                                                 actaconv=False,
                                                                 partida_nac=False,
                                                                 pre=False,
                                                                 observaciones_pre='',
                                                                 fotos=False)
                            documentos.save(request)
                            preguntasinscripcion = inscripcion.preguntas_inscripcion()
                            inscripcion.persona.mi_perfil()
                            inscripciontesdrive = InscripcionTesDrive(inscripcion=inscripcion,
                                                                      licencia=False,
                                                                      record=False,
                                                                      certificado_tipo_sangre=False,
                                                                      prueba_psicosensometrica=False,
                                                                      certificado_estudios=False)
                            inscripciontesdrive.save(request)
                            # inscripcion.mi_malla()
                            inscripcion.malla_inscripcion()
                            inscripcion.actualizar_nivel()
                            if USA_TIPOS_INSCRIPCIONES:
                                inscripciontipoinscripcion = InscripcionTipoInscripcion(inscripcion=inscripcion,
                                                                                        tipoinscripcion_id=TIPO_INSCRIPCION_INICIAL)
                                inscripciontipoinscripcion.save(request)
                            log(u'Adiciono inscripcion: %s' % inscripcion, request, "add")
                            log(u'Importo inscripcion: %s' % inscripcion, request, "add")
                            # persona.creacion_persona(request.session['nombresistema'])
                            linea += 1
                            print(linea)
                            transaction.savepoint_commit(puntosalva)
                        except Exception as ex:
                            transaction.savepoint_rollback(puntosalva)
                            return JsonResponse({"result": "bad", "mensaje": u"Error al ingresar la linea: %s" % linea})
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'fechainicioconvalidacion':
            try:
                form = FechaInicioConvalidacionInscripcionForm(request.POST)
                inscripcion = Inscripcion.objects.get(pk=request.POST['id'])
                if form.is_valid():
                    inscripcion.fechainicioconvalidacion = form.cleaned_data['fecha']
                    inscripcion.save(request)
                    log(u'Adiciono fecha inicio convalidacion: %s' % inscripcion, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'fechainicioprimernivel':
            try:
                form = FechaInicioPrimerNivelInscripcionForm(request.POST)
                inscripcion = Inscripcion.objects.get(pk=int(request.POST['id']))
                if form.is_valid():
                    inscripcion.fechainicioprimernivel = form.cleaned_data['fecha']
                    inscripcion.save(request)
                    log(u'Adiciono fecha inicio primer nivel: %s' % inscripcion, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'fechainiciocarrera':
            try:
                form = FechaInicioCarreraInscripcionForm(request.POST)
                inscripcion = Inscripcion.objects.get(pk=int(request.POST['id']))
                if form.is_valid():
                    inscripcion.fechainiciocarrera = form.cleaned_data['fecha']
                    inscripcion.save(request)
                    log(u'Adiciono fecha inicio carrera: %s' % inscripcion, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'deldocumento':
            try:
                archivo = Archivo.objects.get(pk=request.POST['id'])
                inscripcion = archivo.inscripcion
                archivo.delete()
                log(u'Elimino documento de inscripcion: %s - %s' % (inscripcion, archivo), request, "del")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'asignaturas':
            try:
                inscripcion = Inscripcion.objects.get(pk=request.POST['id'])
                if int(request.POST['tipo']) == 1:
                    # if request.user.has_perm('sga.puede_modificar_ingles'):
                    #     idasignatura = AsignaturaMalla.objects.values_list('asignatura__id', flat=True).filter(malla__carrera__id=34, status=True)
                    #     asignaturas = Asignatura.objects.filter(status=True, id__in=idasignatura)
                    # else:
                    idasignatura = AsignaturaMalla.objects.values_list('asignatura__id', flat=True).filter(status=True, malla__carrera__id=34)
                    asignaturas = Asignatura.objects.filter(status=True, id__in=[x.asignatura.id for x in inscripcion.malla_inscripcion().malla.asignaturamalla_set.all()]).exclude(id__in=idasignatura)
                # asignaturas = Asignatura.objects.filter(id__in=[x.asignatura.id for x in inscripcion.malla_inscripcion().malla.asignaturamalla_set.all()])
                else:
                    # if request.user.has_perm('sga.puede_modificar_ingles'):
                    #     asignaturas = Asignatura.objects.filter(status=True, asignaturamalla__malla__carrera__id=34)
                    # else:
                    asignaturas = Asignatura.objects.filter(status=True).exclude(asignaturamalla__malla__carrera__id=34)
                # asignaturas = Asignatura.objects.all()
                return JsonResponse({"result": "ok", 'listado': [(x.id, x.nombre) for x in asignaturas]})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'activar':
            try:
                inscripcion = Inscripcion.objects.get(pk=request.POST['id'])
                ui = inscripcion.persona.usuario
                ui.is_active = True
                ui.save()
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'desactivar':
            try:
                inscripcion = Inscripcion.objects.get(pk=request.POST['id'])
                ui = inscripcion.persona.usuario
                ui.is_active = False
                ui.save()
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'desactivarperfil':
            try:
                inscripcion = Inscripcion.objects.get(pk=request.POST['id'])
                inscripcion.activo = False
                inscripcion.save(request)
                log(u'Desactivo perfil de usuario: %s' % inscripcion, request, "edit")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'bloqueomatricula':
            try:
                matricula = Matricula.objects.get(pk=request.POST['id'])
                rubrosvencidos=None
                if matricula.inscripcion.carrera.mi_coordinacion2() == 7:
                    cnmoodle = connections['moodle_pos'].cursor()
                    accionretiro = matricula.bloqueomatricula
                    hoy = datetime.now().date()
                    rubrosvencidos = matricula.rubro_set.filter(fechavence__lt=hoy, tipo__tiporubro=1, cancelado=False, status=True).count()
                    if accionretiro == True:
                        # Obtener total rubros vencidos

                        if rubrosvencidos < 2:
                            # Desbloqueo de matrícula
                            matricula.bloqueomatricula = False
                            matricula.save()

                            # Consulto usuario de moodle posgrado
                            usermoodle = matricula.inscripcion.persona.idusermoodleposgrado

                            if usermoodle != 0:
                                # Consulta en mooc_user
                                sql = """Select id, username From mooc_user Where deleted=1 and id=%s""" % (usermoodle)
                                cnmoodle.execute(sql)
                                registro = cnmoodle.fetchall()
                                if registro:
                                    idusuario = registro[0][0]
                                    username = registro[0][1]

                                    # Asignar estado deleted = 0 para que pueda acceder al aula virtual
                                    sql = """Update mooc_user Set deleted=0 Where id=%s""" % (idusuario)
                                    cnmoodle.execute(sql)
                    else:
                        if rubrosvencidos >= 2:
                            # Desbloqueo de matrícula
                            matricula.bloqueomatricula = True
                            matricula.save()

                            # Consulto usuario de moodle posgrado
                            usermoodle = matricula.inscripcion.persona.idusermoodleposgrado

                            if usermoodle != 0:
                                # Consulta en mooc_user
                                sql = """Select id, username From mooc_user Where deleted=0 and id=%s""" % (usermoodle)
                                cnmoodle.execute(sql)
                                registro = cnmoodle.fetchall()
                                if registro:
                                    idusuario = registro[0][0]
                                    username = registro[0][1]

                                    # Asignar estado deleted = 0 para que pueda acceder al aula virtual
                                    sql = """Update mooc_user Set deleted=1 Where id=%s""" % (idusuario)
                                    cnmoodle.execute(sql)
                else:
                    usermoodle = matricula.inscripcion.persona.usuario.username
                    accionretiro = matricula.bloqueomatricula
                    matricula.bloqueomatricula = not accionretiro
                    matricula.save(request)
                    if matricula.inscripcion.carrera.mi_coordinacion2() == 9:
                        cnmoodle = connections['db_moodle_virtual'].cursor()
                    else:
                        cnmoodle = connections['moodle_db'].cursor()
                    if usermoodle and accionretiro:
                        # Consulta en mooc_user
                        sql = """Select id From mooc_user Where suspended=1 and username='%s'""" % (usermoodle)
                        cnmoodle.execute(sql)
                        registro = cnmoodle.fetchall()
                        if registro:
                            sql = """Update mooc_user Set suspended=0 Where username='%s'""" % (usermoodle)
                            cnmoodle.execute(sql)
                log(u'%s: %s' % (u'Desbloquear matricula' if accionretiro else u'Bloquear matricula', matricula), request, "edit")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'activarperfil':
            try:
                inscripcion = Inscripcion.objects.get(pk=request.POST['id'])
                inscripcion.activo = True
                inscripcion.save(request)
                log(u'Activo perfil de usuario: %s' % inscripcion, request, "edit")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'buscarpreinscripcion':
            try:
                if Persona.objects.filter(Q(cedula=request.POST['ced']) | Q(pasaporte=request.POST['ced'])).exists():
                    personains = Persona.objects.filter(Q(cedula=request.POST['ced']) | Q(pasaporte=request.POST['ced']))[0]
                    if not personains.usuario:
                        username = calculate_username(personains)
                        generar_usuario(personains, username, ALUMNOS_GROUP_ID)
                if PreInscrito.objects.filter(cedula=request.POST['ced']).exists():
                    preinscrito = PreInscrito.objects.filter(cedula=request.POST['ced'])[0]
                    return JsonResponse({'result': 'ok', 'preinscrito': unicode(preinscrito), 'preinscrito_id': preinscrito.id})
                else:
                    return JsonResponse({'result': 'bad'})
            except Exception as ex:
                return JsonResponse({'result': 'bad', "mensaje": u'Error al obtener los datos'})

        elif action == 'buscarpersona':
            try:
                if Persona.objects.filter(Q(cedula=request.POST['ced']) | Q(pasaporte=request.POST['ced'])).exists():
                    persona = Persona.objects.filter(Q(cedula=request.POST['ced']) | Q(pasaporte=request.POST['ced']))[0]
                    data_ = model_to_dict(persona,exclude=['archivoplanillaluz','archivocroquis'])
                    return JsonResponse({'result': 'ok','persona_':data_, 'apellido1': persona.apellido1, 'apellido2': persona.apellido2, 'nombres': persona.nombres, 'sexo': persona.sexo_id})
                else:
                    return JsonResponse({'result': 'bad'})
            except Exception as ex:
                return JsonResponse({'result': 'bad', "mensaje": u'Error al obtener los datos'})

        elif action == 'resetear':
            try:
                inscripcion = Inscripcion.objects.get(pk=request.POST['id'])
                resetear_clave(inscripcion.persona)
                log(u'Reseteo clave de inscripcion: %s' % inscripcion, request, "add")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'aplicab2':
            try:
                inscripcion = Inscripcion.objects.get(pk=request.POST['id'])
                inscripcion.aplica_b2=request.POST['valor']
                inscripcion.save(request)
                log(u'Aplico o Quito B2: %s' % inscripcion, request, "add")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'retirocarrera':
            try:
                inscripcion = Inscripcion.objects.get(pk=request.POST['id'])
                form = RetiradoMateriaForm(request.POST)
                if form.is_valid():
                    if not inscripcion.retirocarrera_set.exists():
                        retiro = RetiroCarrera(inscripcion=inscripcion,
                                               fecha=datetime.now().date(),
                                               motivo=form.cleaned_data['motivo'])
                        retiro.save(request)
                    log(u'Retiro de carrera: %s' % inscripcion, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'addadministrativo':
            try:
                inscripcion = Inscripcion.objects.get(pk=request.POST['id'])
                if inscripcion.persona.es_administrador():
                    return JsonResponse({"result": "bad", "mensaje": u"Ya existe un perfil administrativo para este usuario."})
                administrativo = Administrativo(persona=inscripcion.persona,
                                                contrato='',
                                                fechaingreso=datetime.now().date())
                administrativo.save(request)
                g = Group.objects.get(pk=variable_valor('ADMINISTRATIVOS_GROUP_ID'))
                g.user_set.add(inscripcion.persona.usuario)
                g.save()
                inscripcion.persona.crear_perfil(administrativo=administrativo)
                log(u'Adiciono personal administrativo: %s' % administrativo, request, "add")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'adddocente':
            try:
                inscripcion = Inscripcion.objects.get(pk=request.POST['id'])
                if inscripcion.persona.es_profesor():
                    return JsonResponse({"result": "bad", "mensaje": u"Ya existe un perfil de docente para este usuario."})
                profesor = Profesor(persona=inscripcion.persona,
                                    activo=True,
                                    fechaingreso=datetime.now().date(),
                                    coordinacion=Coordinacion.objects.all()[0],
                                    dedicacion=TiempoDedicacionDocente.objects.all()[0])
                profesor.save(request)
                grupo = Group.objects.get(pk=PROFESORES_GROUP_ID)
                grupo.user_set.add(inscripcion.persona.usuario)
                grupo.save()
                inscripcion.persona.crear_perfil(profesor=profesor)
                log(u'Adiciono profesor: %s' % profesor, request, "add")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'infoasignatura':
            try:
                inscripcion = Inscripcion.objects.get(pk=request.POST['iid'])
                malla = inscripcion.malla_inscripcion().malla
                if malla.asignaturamalla_set.filter(asignatura__id=request.POST['id']).exists():
                    asignaturamalla = malla.asignaturamalla_set.filter(asignatura__id=request.POST['id'])[0]
                    return JsonResponse({'result': 'ok', 'creditos': asignaturamalla.creditos, 'horas': asignaturamalla.horas})
                else:
                    asignatura = Asignatura.objects.get(pk=request.POST['id'])
                    return JsonResponse({'result': 'ok', 'creditos': asignatura.creditos, 'horas': 0})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'practicapreprofecionales':
            try:
                data['inscripcion'] = Inscripcion.objects.get(pk=int(request.POST['id']))
                data['practicas'] = PracticasPreprofesionalesInscripcion.objects.filter(culminada=True, status=True,inscripcion_id=int(request.POST['id']))
                template = get_template("inscripciones/practicaspreprofecionales.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'data': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'requisitostitulacion':
            try:
                data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=int(request.POST['id']))
                data['malla'] = malla = inscripcion.mi_malla()
                data['requisitostitulacion'] = requisitotitulacion = malla.requisitotitulacionmalla_set.filter(status=True)
                lista = []
                for lis in requisitotitulacion:
                    valida = False
                    vistafuncion = ''
                    for li in CHOICES_FUNCION_VIEW_REQUISITO:
                        if lis.requisito.funcion == li[0]:
                            valida = True
                    if valida:
                        if dict(CHOICES_FUNCION_VIEW_REQUISITO)[lis.requisito.funcion]:
                            vistafuncion = dict(CHOICES_FUNCION_VIEW_REQUISITO)[lis.requisito.funcion]
                    lista.append([lis.requisito.nombre, lis.run(inscripcion.id), lis.detarequisitos(inscripcion.id), vistafuncion])
                ultimonivelmalla = 0
                penultimonivel = 0
                if inscripcion.mi_malla():
                    ultimonivelmalla = inscripcion.mi_malla().ultimo_nivel_malla()
                    penultimonivel = inscripcion.mi_malla().ultimo_nivel_malla().orden - 1
                data['ultimonivelmalla'] = ultimonivelmalla
                data['penultimonivel'] = penultimonivel
                data['listrequisitos'] = lista
                template = get_template("inscripciones/requisitostitulacion.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'data': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'proyectovinculacion':
            try:
                data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=int(request.POST['id']))
                data['vinculaciones'] = inscripcion.mis_proyectos_vinculacion()
                template = get_template("inscripciones/proyectovinculacion.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'data': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'listar_autorizar':
            try:
                data['matricula'] = matricula = Matricula.objects.get(pk=int(request.POST['idi']))
                data['autorizados'] = autorizar = AutorizarAlumnoSolicitud.objects.filter(matricula=matricula)
                data['nfilas'] = autorizar.count()
                data['form2'] = AutorizarAlumnoSolicitudForm()
                data['hoy'] = datetime.now().date().strftime('%d-%m-%Y')
                template = get_template("inscripciones/autorizarjustificacion.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'data': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'autorizarsol':
            try:
                fini = convertir_fecha(request.POST['fini'])
                ffin = convertir_fecha(request.POST['ffin'])
                if not fini <= ffin:
                    return JsonResponse({"result": "bad", "mensaje": u"La fecha inicio debe ser mayor que la fecha fin"})
                matricula = Matricula.objects.get(pk=int(request.POST['idi']))
                autorizar = AutorizarAlumnoSolicitud(matricula=matricula,
                                                     fechaautorizacion=datetime.now(),
                                                     fechainicio=fini,
                                                     fechafin= ffin)
                autorizar.save(request)
                log(u'Adiciono autorizacion justificación: %s - %s [%s]' % (autorizar.fechainicio, autorizar.fechafin, autorizar.id), request, "add")
                return JsonResponse({"result": "ok", 'id':autorizar.id, 'fechaautorizacion': datetime.now().strftime('%d-%m-%Y'), 'horaautorizacion': datetime.now().strftime("%H:%M"), 'fechainicio': autorizar.fechainicio.strftime('%d-%m-%Y'), 'fechafin': autorizar.fechafin.strftime('%d-%m-%Y')})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'delautorizarsol':
            try:
                autorizar = AutorizarAlumnoSolicitud.objects.get(pk=int(request.POST['id']))
                log(u'Elimino autorizacion justificación: %s - %s [%s]' % (autorizar.fechainicio, autorizar.fechafin, autorizar.id), request, "add")
                autorizar.delete()
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'change_date_naci':
            try:
                persona = Persona.objects.get(pk=int(request.POST['idp']))
                fecha = datetime.strptime(request.POST['fecha'], "%d-%m-%Y")
                persona.nacimiento = fecha
                persona.save(request)
                log(u'Modifico fecha de nacimiento: %s' % persona, request, "add")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'consultafecha':
            try:
                if 'id' in request.POST:
                    inscripcion = Inscripcion.objects.get(pk=int(request.POST['id']))
                    return JsonResponse({"result": "ok", "fecha": inscripcion.persona.nacimiento})
                else:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'guardarfechanacimiento':
            try:
                if 'id' in request.POST and 'fecha' in request.POST:
                    inscripcion = Inscripcion.objects.get(pk=int(request.POST['id']))
                    persona = Persona.objects.get(pk=inscripcion.persona.id)
                    fecha = datetime.strptime(request.POST['fecha'], "%d-%m-%Y")
                    persona.nacimiento = fecha
                    persona.save(request)
                    log(u'Modifico fecha de nacimiento: %s' % persona, request, "add")
                    return JsonResponse({"result": "ok", "id": inscripcion.id, "fecha": persona.nacimiento.strftime("%d-%m-%Y")})
                else:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'editmallahistorica':
            try:
                recordacademico = RecordAcademico.objects.get(pk=request.POST['id'])
                f = MallaHistoricaForm(request.POST)
                if f.is_valid():
                    recordacademico.asignaturamallahistorico = f.cleaned_data['asignaturamallahistorico']
                    recordacademico.asignaturamalla = recordacademico.inscripcion.asignatura_en_asignaturamalla(f.cleaned_data['asignaturamallahistorico'].asignatura )
                    recordacademico.save(request)
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'eliminarmasivo':
            try:
                datos = json.loads(request.POST['lista'])
                if not datos:
                    return JsonResponse({"result": "bad", "mensaje": u"Debe seleccionar datos"})
                for elemento in datos:
                    record = RecordAcademico.objects.get(pk=int(elemento['id']))
                    asignatura = record.asignatura
                    inscripcion = record.inscripcion
                    homologacion = record.homologacioninscripcion_set.all()
                    homologacion.delete()
                    convalidacion = record.convalidacioninscripcion_set.all()
                    convalidacion.delete()
                    historico = record.historicorecordacademico_set.all()
                    historico.delete()
                    record.delete()
                    inscripcion.actualizar_nivel()
                    inscripcion.actualiza_matriculas(asignatura)
                log(u'Elimino materias de record: %s - %s' % (personasesion, record.inscripcion.persona), request, "add")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'recordhistorico':
            try:
                from itertools import chain
                data['inscripcion'] = inscripcion = Inscripcion.objects.get(id=request.POST['id'],status=True)
                data['total_creditos_malla'] =total_creditos_malla= inscripcion.total_creditos_malla_historicas()
                data['total_horas'] = total_horas= inscripcion.total_horas_malla_historicas()
                data['promedio'] = promedio = inscripcion.promedio_record()
                data['aprobadas'] = aprobadas = inscripcion.recordacademico_set.filter(aprobada=True, valida=True,asignaturamalla__isnull=False, noaplica=False).count()
                # niveleshistorico = inscripcion.recordacademico_set.values_list('asignaturamallahistorico__nivelmalla', flat=True).filter(status=True, noaplica=False).order_by('asignaturamallahistorico')
                # data['records'] = record = inscripcion.recordacademico_set.filter(status=True, noaplica=False).order_by('asignaturamalla', 'asignaturamallahistorico')
                niveles = []
                nivel=None
                for registro in inscripcion.recordacademico_set.filter(status=True, noaplica=False).order_by('asignaturamalla__nivelmalla', 'asignaturamalla', 'asignaturamallahistorico'):
                    if registro.asignaturamalla:
                        nivel = registro.asignaturamalla.nivelmalla
                    elif registro.asignaturamallahistorico:
                        nivel = registro.asignaturamallahistorico.nivelmalla
                    if not nivel in niveles and not nivel == None:
                        niveles.append(nivel)
                return conviert_html_to_pdf('inscripciones/recordhistorico.html',
                                            {'pagesize': 'A4',
                                             'inscripcion':inscripcion,
                                             'total_creditos_malla':null_to_decimal(total_creditos_malla,2),
                                             'total_horas':total_horas,
                                             'promedio':promedio,
                                             'aprobadas':aprobadas,
                                             'personasesion':personasesion,
                                             'hoy':datetime.now().date(),'niveles':niveles
                                             })
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'cambiarbloqueomatricula':
            try:
                inscripcion = Inscripcion.objects.get(pk=request.POST['id'])
                ant=inscripcion.bloqueomatricula
                if inscripcion.bloqueomatricula:
                    inscripcion.bloqueomatricula=False
                else:
                    inscripcion.bloqueomatricula=True
                inscripcion.save(request)
                log(u'Cambio estado de bloqueo matricula: %s de %s a %s' % (inscripcion,ant,inscripcion.bloqueomatricula), request, "edit")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'desactivarperfilusuario':
            try:
                inscripcion = Inscripcion.objects.get(pk=request.POST['id'])
                persona = inscripcion.persona
                persona.perfilusuario_set.filter(inscripcion=inscripcion).update(visible=False)
                log(u'Desactivo perfil usuario: %s' % (inscripcion), request, "edit")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'activarperfilusuario':
            try:
                inscripcion = Inscripcion.objects.get(pk=request.POST['id'])
                persona = inscripcion.persona
                persona.perfilusuario_set.filter(inscripcion=inscripcion).update(visible=True)
                log(u'Activo perfil usuario: %s' % (inscripcion), request, "edit")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'desactivarsoporte':
            try:
                if VirtualSoporteUsuarioInscripcion.objects.filter(id=request.POST['id'], status=True).exists():
                    soporte=VirtualSoporteUsuarioInscripcion.objects.get(id=request.POST['id'], status=True)
                    soporte.activo=False
                    soporte.save(request)
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": "Error al guardar los datos."})

        elif action == 'activarsoporte':
            try:
                soporte = VirtualSoporteUsuario.objects.get(id=request.POST['soporte'])
                if 'matricula' in request.POST:
                    matricula = Matricula.objects.get(id=request.POST['matricula'])
                    if VirtualSoporteUsuarioInscripcion.objects.filter(status=True, matricula=matricula, soporteusuario=soporte).exists():
                        soporte = VirtualSoporteUsuarioInscripcion.objects.filter(status=True, matricula=matricula, soporteusuario=soporte)[0]
                        if soporte.activo == False:
                            soporte.activo = True
                        else:
                            soporte.activo = False
                    else:
                        soporte= VirtualSoporteUsuarioInscripcion(matricula=matricula, soporteusuario=soporte, activo=True )
                    if soporte:
                        soporte.save(request)
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": "Error al eliminar los datos."})

        elif action == 'resetear_clave_pregrado_virtual':
            try:
                inscripcion = Inscripcion.objects.get(pk=request.POST['id'])
                resetear_clave_pregrado_virtual(inscripcion.persona)
                log(u'Reseteo clave de inscripcion virtual: %s' % inscripcion, request, "add")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'consultar_disponible_usuario':
            try:
                inscripcion = Inscripcion.objects.get(pk=int(encrypt(request.POST['id'])))
                usuario123=request.POST['usuario']
                if User.objects.values('id').filter(username=usuario123).exclude(persona=inscripcion.persona).exists():
                    return JsonResponse({'result': 'nodisponible'})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'editusuario':
            try:
                inscripcion = Inscripcion.objects.get(pk=int(encrypt(request.POST['id'])))
                usuarioaguardar=inscripcion.persona.usuario
                f = NombreUsuarioForm(request.POST)
                if f.is_valid():
                    usuario123 = f.cleaned_data['nombre']
                    if usuarioaguardar.username != usuario123:
                        if not User.objects.values('id').filter(username=usuario123).exists():
                            inscripcion.persona.emailinst="%s@unemi.edu.ec" % usuario123
                            inscripcion.persona.save(request)
                            usuarioaguardar.username=usuario123
                            usuarioaguardar.save()
                            log(u'Modifica nombre de usuario: %s - %s' % (inscripcion.persona, usuario123), request, "edit")
                            return JsonResponse({"result": "ok"})
                        else:
                            JsonResponse({"result": "bad", "mensaje": u"El usuario ingresado no se encuentra disponible."})
                    else:
                        return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'infoperdidagratuidad':
            try:
                inscripcion = Inscripcion.objects.get(pk=int(request.POST['id']))
                data['inscripcion'] = inscripcion
                template = get_template('inscripciones/infoperdidagratuidad.html')
                json_contenido = template.render(data)
                return JsonResponse({"result": "ok", "contenido":json_contenido })
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al consultar los datos."})

        elif action == 'infotarjetacademica':
            try:
                inscripcion = Inscripcion.objects.get(pk=int(request.POST['id']))
                data['inscripcion'] = inscripcion
                data['registros'] = TarjetaRegistroAcademico.objects.filter(inscripcion=inscripcion)
                template = get_template('inscripciones/vertarjetaacademica.html')
                json_contenido = template.render(data)
                return JsonResponse({"result": "ok", "contenido":json_contenido })
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al consultar los datos."})

        elif action == 'auditoria':
            try:
                baseDate = datetime.today()
                year = request.POST['year'] if 'year' in request.POST and request.POST['year'] else baseDate.year
                month = request.POST['month'] if 'month' in request.POST and request.POST['month'] else baseDate.month
                data['idi'] = request.POST['id']
                data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=int(encrypt(request.POST['id'])))
                logs = LogEntry.objects.filter(Q(change_message__icontains=inscripcion.persona.__str__()) | Q(user=inscripcion.persona.usuario), action_time__year=year).exclude(user__is_superuser=True)
                logs1 = LogEntryBackup.objects.filter(Q(change_message__icontains=inscripcion.persona.__str__()) | Q(user=inscripcion.persona.usuario), action_time__year=year).exclude(user__is_superuser=True)
                logs2 = LogEntryBackupdos.objects.filter(Q(change_message__icontains=inscripcion.persona.__str__()) | Q(user=inscripcion.persona.usuario), action_time__year=year).exclude(user__is_superuser=True)
                logs3 = LogEntryLogin.objects.filter(user=inscripcion.persona.usuario, action_time__year=year).exclude(user__is_superuser=True, action_app=2)
                addmaterias = AgregacionEliminacionMaterias.objects.filter(matricula__inscripcion=inscripcion, fecha__year=year).order_by('-fecha')
                if int(month):
                    logs = logs.filter(action_time__month=month)
                    logs1 = logs1.filter(action_time__month=month)
                    logs2 = logs2.filter(action_time__month=month)
                    logs3 = logs3.filter(action_time__month=month)
                    addmaterias = addmaterias.filter(fecha__month=month)
                logslist0 = list(logs.values_list("action_time", "action_flag", "change_message", "user__username"))
                logslist1 = list(logs1.values_list("action_time", "action_flag", "change_message", "user__username"))
                logslist2 = list(logs2.values_list("action_time", "action_flag", "change_message", "user__username"))
                logslist=logslist0+logslist1+logslist2
                aLogList = []
                for xItem in logslist:
                    #print(xItem)
                    if xItem[1] == 1:
                        action_flag = '<label class="label label-success">AGREGAR</label>'
                    elif xItem[1] == 2:
                        action_flag = '<label class="label label-info">EDITAR</label>'
                    elif xItem[1] == 3:
                        action_flag = '<label class="label label-important">ELIMINAR</label>'
                    else:
                        action_flag = '<label class="label label-warning">OTRO</label>'
                    aLogList.append({"action_time": xItem[0],
                                     "action_flag": action_flag,
                                     "change_message": xItem[2],
                                     "username": xItem[3]})
                for xItem in list(logs3.values_list("action_time", "action_flag", "change_message", "user__username", "id")):
                    l = LogEntryLogin.objects.get(pk=xItem[4])
                    if xItem[1] == 1:
                        action_flag = '<label class="label label-success">EXITOSO</label>'
                    elif xItem[1] == 2:
                        action_flag = '<label class="label label-warning">FALLIDO</label>'
                    else:
                        action_flag = '<label class="label label-important">DESCONOCIDO</label>'
                    aLogList.append({"action_time": xItem[0],
                                     "action_flag": action_flag,
                                     "change_message": l.get_data_message(),
                                     "username": xItem[3]
                                     })
                addmateriaslist = list(addmaterias.values_list("fecha", "agregacion", "asignatura__nombre", "responsable__usuario__username"))
                aLogAddMateriaslist = []
                my_time = datetime.min.time()
                for xItem in addmateriaslist:
                    #print(xItem)
                    aLogAddMateriaslist.append({"action_time": datetime.combine(xItem[0], my_time),
                                                "action_flag": ADDITION if xItem[1] else DELETION,
                                                "change_message": u"%s la asignatura %s" % (("Agrego" if xItem[1] else "Elimino"), xItem[2]),
                                                "username": xItem[3]})
                datalogs = aLogList + aLogAddMateriaslist
                data['logs'] = sorted(datalogs, key=lambda x: x['action_time'], reverse=True)
                numYear = 6
                dateListYear = []
                for x in range(0, numYear):
                    dateListYear.append((baseDate.year)-x)
                data['list_years'] = dateListYear
                data['year_now'] = int(year)
                data['month_now'] = int(month)
                template = get_template('inscripciones/auditoria.html')
                json_contenido = template.render(data)
                return JsonResponse({"result": "ok", "contenido":json_contenido })
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al consultar los datos."})

        elif action == 'descargarauditoria':
            try:
                inscripcion = Inscripcion.objects.get(pk=int(encrypt(request.POST['id'])))
                titulo = 'Generando reporte de auditoria'
                noti = Notificacion(cuerpo='Reporte de auditoria en progreso',
                                    titulo=titulo, destinatario=personasesion,
                                    url='',
                                    prioridad=1, app_label='sga-sagest',
                                    fecha_hora_visible=datetime.now() + timedelta(days=1), tipo=2,
                                    en_proceso=True)
                noti.save(request)
                data['persona'] = inscripcion.persona
                reporte_auditoria_background(request=request, data=data, notif=noti.pk).start()
                return JsonResponse({'result':False, 'mensaje':titulo})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result':True, 'mensaje':f'Error: {ex}'})

        elif action == 'delete-cache':
            try:
                try:
                    eInscripcion = Inscripcion.objects.get(pk=int(encrypt(request.POST['id'])))
                except ObjectDoesNotExist:
                    raise NameError(u"No se encontro inscripción")
                ePersona = eInscripcion.persona
                ePersona.delete_cache()
                eInscripcion.delete_cache()
                for eMatricula in Matricula.objects.filter(inscripcion=eInscripcion):
                    eMatricula.delete_cache()
                return JsonResponse({"result": "ok", "mensaje": u""})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": ex.__str__()})

        elif action == 'addppl':
            try:
                f = PersonaPPLForm(request.POST, request.FILES)
                archivoppl = None
                if 'archivoppl' in request.FILES:
                    arch = request.FILES['archivoppl']
                    extension = arch._name.split('.')
                    tam = len(extension)
                    exte = extension[tam - 1]
                    if arch.size > 4194304:
                        raise NameError(u"Tamaño del archivo es mayor a 4 Mb")
                    if not exte.lower() == 'pdf':
                        raise NameError(u"Solo se permiten archivos .pdf")
                    archivoppl = request.FILES['archivoppl']
                    archivoppl._name = generar_nombre("archivoppl_", archivoppl._name)
                if not 'idi' in request.POST:
                    raise NameError(u"Inscripción no encontrada")
                if not Inscripcion.objects.filter(pk=int(request.POST['idi'])).exists():
                    raise NameError(u"Inscripción no encontrada")
                inscripcion = Inscripcion.objects.get(pk=int(request.POST['idi']))

                if not f.is_valid():
                    raise NameError('Formulario incorrecto')

                if HistorialPersonaPPL.objects.filter(persona=inscripcion.persona, fechaingreso=f.cleaned_data['fechaingresoppl']).exists():
                    raise NameError(u"Registro ya existe")

                historialppl = HistorialPersonaPPL(persona=inscripcion.persona,
                                                   observacion=f.cleaned_data['observacionppl'] if f.cleaned_data['observacionppl'] else None,
                                                   archivo=archivoppl,
                                                   fechaingreso=f.cleaned_data['fechaingresoppl'],
                                                   fechasalida=f.cleaned_data['fechasalidappl'] if f.cleaned_data['fechasalidappl'] else None,
                                                   centrorehabilitacion=f.cleaned_data['centrorehabilitacion'] if f.cleaned_data['centrorehabilitacion'] else None,
                                                   lidereducativo=f.cleaned_data['lidereducativo'] if f.cleaned_data['lidereducativo'] else None,
                                                   correolidereducativo=f.cleaned_data['correolidereducativo'] if f.cleaned_data['correolidereducativo'] else None,
                                                   telefonolidereducativo=f.cleaned_data['telefonolidereducativo'] if f.cleaned_data['telefonolidereducativo'] else None,
                                                   )
                historialppl.save(request)
                log(u'Adiciono registro PPL al estudiante: %s' % historialppl, request, "add")
                messages.add_message(request, messages.SUCCESS, f'Se guardo correctamente el registro')
                return JsonResponse({"result": "ok"})

            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos. %s" % ex.__str__()})

        elif action == 'editppl':
            try:
                if not HistorialPersonaPPL.objects.filter(pk=int(request.POST['id'])).exists():
                    raise NameError(u"Registro no existe")
                hppl = HistorialPersonaPPL.objects.get(pk=int(request.POST['id']))
                archivoppl = None
                f = PersonaPPLForm(request.POST, request.FILES)
                if 'archivoppl' in request.FILES:
                    arch = request.FILES['archivoppl']
                    extension = arch._name.split('.')
                    tam = len(extension)
                    exte = extension[tam - 1]
                    if arch.size > 4194304:
                        raise NameError(u"Tamaño del archivo es mayor a 4 Mb")
                    if not exte.lower() == 'pdf':
                        raise NameError(u"Solo se permiten archivos .pdf")
                    archivoppl = request.FILES['archivoppl']
                    archivoppl._name = generar_nombre("archivoppl_", archivoppl._name)

                if not f.is_valid():
                    raise NameError('Formulario incorrecto')

                if HistorialPersonaPPL.objects.filter(persona=hppl.persona, fechaingreso=f.cleaned_data['fechaingresoppl']).exclude(pk=hppl.id).exists():
                    raise NameError(u"Registro ya existe")
                hppl.fechaingreso =f.cleaned_data['fechaingresoppl']
                hppl.fechasalida = f.cleaned_data['fechasalidappl'] if f.cleaned_data['fechasalidappl'] else None
                hppl.centrorehabilitacion = f.cleaned_data['centrorehabilitacion'] if f.cleaned_data['centrorehabilitacion'] else None
                hppl.lidereducativo = f.cleaned_data['lidereducativo'] if f.cleaned_data['lidereducativo'] else None
                hppl.correolidereducativo = f.cleaned_data['correolidereducativo'] if f.cleaned_data['correolidereducativo'] else None
                hppl.telefonolidereducativo = f.cleaned_data['telefonolidereducativo'] if f.cleaned_data['telefonolidereducativo'] else None
                if archivoppl:
                    hppl.archivo = archivoppl
                hppl.observacion = f.cleaned_data['observacionppl'] if f.cleaned_data['observacionppl'] else None
                hppl.save(request)
                log(u'Edito registro PPL al estudiante: %s' % hppl, request, "edit")
                messages.add_message(request, messages.SUCCESS, f'Se guardo correctamente el registro')
                return JsonResponse({"result": "ok"})

            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos. %s" % ex.__str__()})

        elif action == 'actualizar_nivel_malla':
            try:
                inscripcion = Inscripcion.objects.get(pk=int(encrypt(request.POST['id'])))
                data['inscripcion'] = inscripcion
                inscripcion.actualizar_nivel()
                log(u'Actualizo Nivel Malla: %s' % inscripcion, request, "edit")
                return JsonResponse({"result": "ok", "mensaje": u'Se Actualizo correctamente el Nivel Malla'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al consultar los datos."})

        elif action == 'addsedeexamen':
            try:
                from inno.models import MatriculaSedeExamen
                if not 'id' in request.POST:
                    raise NameError(u"Matrícula no encontrada")
                eMatricula = Matricula.objects.filter(pk=int(request.POST['id'])).first()
                sede = request.POST['sede_id']
                detallemodeloevaluativo = request.POST['modelo']
                if not eMatricula:
                    raise NameError(u"Matrícula no encontrada")
                if not sede:
                    raise NameError(u"Debe seleccionar una sede")
                if not detallemodeloevaluativo:
                    raise NameError(u"Debe seleccionar un modelo evaluativo")
                planificacion = MatriculaSedeExamen.objects.filter(matricula=eMatricula, sede_id=sede, detallemodeloevaluativo_id=detallemodeloevaluativo).first()
                if planificacion:
                    raise NameError(u"El registro ya existe")
                eMatriculaSedeExamen = MatriculaSedeExamen(matricula=eMatricula,sede_id=sede, detallemodeloevaluativo_id=detallemodeloevaluativo)
                eMatriculaSedeExamen.save(request)
                log(u'Adiciono sede de examen: %s' % eMatriculaSedeExamen, request, "add")
                # messages.add_message(request, messages.SUCCESS, f'Se guardo correctamente el registro')
                return JsonResponse({"result": "ok", 'mensaje': 'Se guardó correctamente el registro'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos. %s" % ex.__str__()})

        elif action == 'editsedeexamen':
            try:
                from inno.models import MatriculaSedeExamen
                if not 'matsede' in request.POST:
                    raise NameError(u"Registro no encontrado")
                sede = request.POST['sede_id']
                detallemodeloevaluativo = request.POST['modelo']
                eMatriculaSedeExamen = MatriculaSedeExamen.objects.filter(pk=int(request.POST['matsede'])).first()
                if not eMatriculaSedeExamen:
                    raise NameError(u"Registro no encontrada")
                eMatricula = eMatriculaSedeExamen.matricula
                if MatriculaSedeExamen.objects.filter(matricula=eMatricula, sede_id=sede, detallemodeloevaluativo_id=detallemodeloevaluativo).exclude(pk=eMatriculaSedeExamen.pk).exists():
                    raise NameError(u"El registro ya existe")
                eMatriculaSedeExamen.sede_id=sede
                eMatriculaSedeExamen.detallemodeloevaluativo_id=detallemodeloevaluativo
                eMatriculaSedeExamen.save(request, update_fields=['detallemodeloevaluativo_id', 'sede_id'])
                log(u'Edito sede de examen: %s' % eMatriculaSedeExamen, request, "edit")
                # messages.add_message(request, messages.SUCCESS, f'Se guardo correctamente el registro')
                return JsonResponse({"result": "ok", 'mensaje': 'Se guardó correctamente el registro'})

            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos. %s" % ex.__str__()})

        elif action == 'deletesedeexamen':
            try:
                from inno.models import MatriculaSedeExamen
                if not 'id' in request.POST:
                    raise NameError(u"Registro no encontrado")
                if not MatriculaSedeExamen.objects.filter(pk=int(request.POST['id'])).exists():
                    raise NameError(u"Registro no encontrada")
                eEliminar = eMatriculaSedeExamen = MatriculaSedeExamen.objects.get(pk=int(request.POST['id']))
                eMatriculaSedeExamen.delete()
                log(u'Elimino sede de examen: %s' % eEliminar, request, "del")
                return JsonResponse({"result": "ok", "mensaje": "Se elimino correctamente el registro"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos. %s" % ex.__str__()})

        elif action == 'deleteAlumnoPlanificacion':
            try:
                if not 'id' in request.POST:
                    raise NameError(u"No se encontro parametro correcto")
                if not MateriaAsignadaPlanificacionSedeVirtualExamen.objects.values("id").filter(pk=request.POST['id']):
                    raise NameError(u"No se encontro la materia a eliminar")
                eMateriaAsignadaPlanificacionSedeVirtualExamen = eDelete = MateriaAsignadaPlanificacionSedeVirtualExamen.objects.get(pk=request.POST['id'])
                if eMateriaAsignadaPlanificacionSedeVirtualExamen.habilitadoexamen:
                    raise NameError(f"Alumno {eMateriaAsignadaPlanificacionSedeVirtualExamen.materiaasignada.matricula.inscripcion.persona} tiene habilitado el examen, no puede eliminarlo")
                eDelete.delete()
                log(u'Elimino planificación de alumno: %s' % eMateriaAsignadaPlanificacionSedeVirtualExamen, request, "del")
                return JsonResponse({"result": True})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": False, 'message': str(ex)})

        elif action == 'obtenerSENESCYT':
            try:
                from soap.consumer.senescyt import Titulos
                if not 'id' in request.POST:
                    raise NameError('No se encontro parametro de inscripción')
                eInscripcion = Inscripcion.objects.get(pk=request.POST['id'])
                ePersona = eInscripcion.persona
                identificacion = ePersona.identificacion()
                eTitulos = Titulos(identificacion)
                mistitulos = eTitulos.consultar()
                eInscripcion.actualiza_estado_matricula()
                return JsonResponse({"result": 'ok', "mensaje": 'Se obtuvo la información correctamente'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": 'bad', "mensaje": f"Ocurrio un error: {ex.__str__()}"})

        elif action == 'upload':
            from inno.models import MatriculaSedeExamen
            try:
                if not 'documento' in request.FILES:
                    raise NameError(f"Por favor carge un documento")
                documento = request.FILES['documento']
                if documento.size > 10485760:
                    raise NameError(u"Archivo mayor a 10 Mb.")
                documentood = documento._name
                ext = documentood[documentood.rfind("."):]
                if not ext in ['.xls', '.xlsx']:
                    raise NameError(u"Solo archivo con extensión. xls, xlsx.")
                if not 'modelo' in request.POST:
                    raise NameError('No eligió un modelo evaluativo')
                if not 'sede' in request.POST:
                    raise NameError('No eligió una sede')
                sede = SedeVirtual.objects.get(id=request.POST['sede'])
                modelo = DetalleModeloEvaluativo.objects.get(id=request.POST['modelo'])
                tipo = int(request.POST['tipo'])
                workbook = openpyxl.load_workbook(documento)
                sheet = workbook.worksheets[0]
                all_rows = sheet.rows
                linea = 0
                col_inscripcion = 1
                col_sede_id = sede.id
                col_cedula = 0
                periodo_id = periodo.id
                detallemodeloevaluativo_id = modelo.id
                eMatriculas = None
                if not all_rows:
                    raise NameError('La matriz está vacia')
                for fila in all_rows:
                    linea += 1
                    if linea > 1:
                        try:
                            inscripcion_id = int(fila[col_inscripcion].value)
                        except Exception as e:
                            inscripcion_id = 0
                        try:
                            cedula = fila[col_cedula].value if fila[col_cedula].value not in [' ', None] else None
                        except Exception as e:
                            cedula = None
                        if inscripcion_id > 0:
                            eMatriculas = Matricula.objects.filter(inscripcion_id=inscripcion_id, status=True, nivel__periodo_id=periodo_id)
                        elif cedula:
                            eMatriculas = Matricula.objects.filter(inscripcion__persona__cedula=cedula, status=True, nivel__periodo_id=periodo_id)
                        # if not eMatriculas:
                        #     raise NameError(u"No existe matriculas con los datos proporcionados")
                        if tipo == 0:
                            eMatriculas = eMatriculas.filter(inscripcion__carrera__coordinacion__lte=5)
                        elif tipo == 1:
                            eMatriculas = eMatriculas.filter(inscripcion__carrera__coordinacion=9)
                        else:
                            eMatriculas = eMatriculas.filter(inscripcion__carrera__coordinacion=7)
                        if not eMatriculas:
                            # raise NameError(u"No existe matriculas con los datos proporcionados")
                            # eMatricula = eMatriculas.first()
                            continue
                        for eMatricula in eMatriculas:
                            with transaction.atomic():
                                try:
                                    eMatriculaSedeExamenes = MatriculaSedeExamen.objects.filter(matricula=eMatricula)
                                    if not eMatriculaSedeExamenes.values("id").exists():
                                        eMatriculaSedeExamen = MatriculaSedeExamen(matricula=eMatricula,
                                                                                   detallemodeloevaluativo=modelo,
                                                                                   sede=sede)
                                        eMatriculaSedeExamen.save()
                                    elif eMatriculaSedeExamenes.filter(detallemodeloevaluativo=modelo, sede=sede).values(
                                            "id").exists():
                                        eMatricula.delete_cache()
                                        eInscripcion = eMatricula.inscripcion
                                        eInscripcion.delete_cache()
                                        ePersona = eInscripcion.persona
                                        ePersona.delete_cache()
                                        continue
                                    else:
                                        eMatriculaSedeExamen = eMatriculaSedeExamenes.first()
                                        eMatriculaSedeExamen.sede = sede
                                        eMatriculaSedeExamen.detallemodeloevaluativo = modelo
                                        eMatriculaSedeExamen.save()
                                        eMatricula.delete_cache()
                                        eInscripcion = eMatricula.inscripcion
                                        eInscripcion.delete_cache()
                                        ePersona = eInscripcion.persona
                                        ePersona.delete_cache()
                                    # print(f"{eMatriculaSedeExamen.matricula.inscripcion.persona} -- {eMatriculaSedeExamen.sede}")
                                except Exception as ex:
                                    transaction.set_rollback(True)
                                    return JsonResponse({"result": "bad", "mensaje": u"Solicitud Incorrecta."})
                return JsonResponse({'result': 'ok'})
            except Exception as e:
                return JsonResponse({"result": "bad", "mensaje": str(e)})

        elif action == 'actionHorarioExamen':
            try:
                id = request.POST.get('id', None)
                if id is None:
                    raise NameError('No se encontro parametro')
                visiblehorarioexamen = request.POST.get('visiblehorarioexamen', 'ocultar')
                visiblehorarioexamen = visiblehorarioexamen == 'mostrar'
                try:
                    eMatricula = Matricula.objects.get(pk=id)
                except ObjectDoesNotExist:
                    raise NameError('No se encontro matricula')
                MateriaAsignada.objects.filter(matricula=eMatricula).update(visiblehorarioexamen=visiblehorarioexamen)
                eMatricula.delete_cache()
                eInscripcion = eMatricula.inscripcion
                eInscripcion.delete_cache()
                ePersona = eInscripcion.persona
                ePersona.delete_cache()
                return JsonResponse({"result": 'ok', "mensaje": f'Se {"activo" if visiblehorarioexamen else "inactivo"} horario de examen correctamente'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": 'bad', "mensaje": f"Ocurrio un error: {ex.__str__()}"})

        elif action == 'actionResetExamen':
            from inno.models import MatriculaSedeExamen
            try:
                id = request.POST.get('id', None)
                if id is None:
                    raise NameError('No se encontro parametro')
                try:
                    eMatriculaSedeExamen = MatriculaSedeExamen.objects.get(pk=id)
                except ObjectDoesNotExist:
                    raise NameError('No se encontro matricula')
                eMatriculaSedeExamen.archivoidentidad = None
                eMatriculaSedeExamen.archivofoto = None
                eMatriculaSedeExamen.aceptotermino = False
                eMatriculaSedeExamen.fechaaceptotermino = None
                eMatriculaSedeExamen.urltermino = None
                eMatriculaSedeExamen.save(request)
                eMatricula = eMatriculaSedeExamen.matricula
                eMatricula.delete_cache()
                eInscripcion = eMatricula.inscripcion
                eInscripcion.delete_cache()
                ePersona = eInscripcion.persona
                ePersona.delete_cache()
                log(u'Restablecio proceso de archivos y acuerdo de terminos y condiciones de examenes finales: %s' % eMatriculaSedeExamen, request, "add")
                return JsonResponse({"result": 'ok', "mensaje": f'Se restablecio proceso de examenes'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": 'bad', "mensaje": f"Ocurrio un error: {ex.__str__()}"})

        elif action == 'mostrarhorarios':
            try:
                accion = request.POST.get('accion', None)
                try:
                    accion = int(accion)
                except:
                    accion = 0
                titulo = f'Proceso de mostrar horarios de pregrado en proceso.'
                datos = {'accion': accion, 'periodo': periodo, 'persona': personasesion, 'es_admision': False, 'es_pregrado': True}
                noti = Notificacion(
                    cuerpo='Se inicializo el proceso de mostar horarios de exámenes de pregrado.',
                    titulo=titulo, destinatario=personasesion,
                    url='',
                    prioridad=1, app_label='SGA',
                    fecha_hora_visible=datetime.now() + timedelta(days=1), tipo=2,
                    en_proceso=True)
                noti.save(request)
                actualizar_visible_horario_masivo(request=request, data=datos, notif=noti.pk).start()
                return JsonResponse({"result": 'ok', "mensaje": f'Proceso de mostrar horarios en proceso de pregrado.'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": 'bad', "mensaje": f"Ocurrio un error: {ex.__str__()}"})

        elif action == 'limpiarcache':
            try:
                titulo = f'Proceso de limpieza de cache de pregrado en proceso.'
                datos = {'periodo': periodo, 'persona': personasesion, 'es_admision': False, 'es_pregrado': True}
                noti = Notificacion(
                    cuerpo='Se inicializo el proceso de limpieza de cache de pregrado.',
                    titulo=titulo,
                    destinatario=personasesion,
                    url='',
                    prioridad=1, app_label='SGA',
                    fecha_hora_visible=datetime.now() + timedelta(days=1), tipo=2,
                    en_proceso=True)
                noti.save(request)
                limpiar_cache_masivo(request=request, data=datos, notif=noti.pk).start()
                return JsonResponse({"result": 'ok', "mensaje": f'Proceso de mostrar horarios en proceso.'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": 'bad', "mensaje": f"Ocurrio un error: {ex.__str__()}"})

        elif action == 'acept_automatricula':
            try:
                id = int(encrypt((request.POST.get('id', encrypt('0')))))
                if id == 0:
                    raise NameError('No se encontro parametro')
                try:
                    eMatricula = Matricula.objects.get(pk=id)
                except ObjectDoesNotExist:
                    raise NameError('No se encontro matricula')
                eMatricula.termino=True
                eMatricula.fechatermino = datetime.now().date()
                eMatricula.save()
                if eMatricula.termino:
                    titulo = f'Proceso de limpieza de cache de pregrado en proceso.'
                    datos = {'periodo': periodo, 'persona': personasesion, 'es_admision': False, 'es_pregrado': True}
                    noti = Notificacion(
                        cuerpo='Se inicializo el proceso de limpieza de cache de pregrado.',
                        titulo=titulo,
                        destinatario=personasesion,
                        url='',
                        prioridad=1, app_label='SGA',
                        fecha_hora_visible=datetime.now() + timedelta(days=1), tipo=2,
                        en_proceso=True)
                    noti.save(request)
                    limpiar_cache_masivo(request=request, data=datos, notif=noti.pk).start()

                    return JsonResponse({"result": 'ok', "mensaje": f'Automatricula Aceptada. Cache Limpiada'})
                else:
                    return JsonResponse({"result": 'bad', "mensaje": f'Ocurrio un error'})


            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": 'bad', "mensaje": f"Ocurrio un error: {ex.__str__()}"})

        elif action == 'savemoveaula':
            try:
                if not 'matid' in request.POST:
                    raise NameError('Materia no encontrada')
                if not 'aula' in request.POST:
                    raise NameError('Aula no encontrada')
                if not 'modelo' in request.POST:
                    raise NameError('Modelo evaluativo no encontrado')
                materiaasignada = MateriaAsignada.objects.filter(status=True, id=request.POST['matid']).first()
                if not materiaasignada:
                    raise NameError('Materia no encontrada')
                aula = AulaPlanificacionSedeVirtualExamen.objects.filter(status=True, id=request.POST['aula']).first()
                if not aula:
                    raise NameError('Aula no encontrada')
                planificacion = materiaasignada.planficacionvirtualexamen(int(request.POST['modelo'])).first()
                if planificacion:
                    planificacion.aulaplanificacion = aula
                    planificacion.password = aula.password
                    planificacion.save(request, update_fields=['aulaplanificacion', 'password'])
                    log(u'Modificó sede de examen: %s' % planificacion, request, "edit")
                else:
                    # raise NameError('Lo sentimos esta materia no se puede mover ya que no está planificada')
                    #detallemodeloevaluativo_id = materiaasignada.modelo
                    planificacion = MateriaAsignadaPlanificacionSedeVirtualExamen(
                        aulaplanificacion=aula,
                        materiaasignada=materiaasignada,
                        detallemodeloevaluativo_id=request.POST['modelo'],
                        utilizar_qr=True,
                        password=aula.password
                    )
                    planificacion.save(request)
                    log(u'Agregó sede de examen: %s' % planificacion, request, "add")

                return JsonResponse({"result": 'ok'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos. %s" % ex.__str__()})

        return JsonResponse({"result": "bad", "mensaje": u"Solicitud Incorrecta."})

    else:
        data['title'] = u'Listado de inscripciones'
        persona = request.session['persona']
        if 'action' in request.GET:
            action = request.GET['action']

            if action == 'cambiomalla':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_inscripciones')
                    data['title'] = u'Cambio de malla'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    puede_modificar_inscripcion(request, inscripcion)
                    form = CambiomallaForm()
                    form.mallas(inscripcion.carrera)
                    data['form'] = form
                    return render(request, "inscripciones/cambiomalla.html", data)
                except Exception as ex:
                    pass

            elif action == 'retirocarrera':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_inscripciones')
                    data['title'] = u'Retiro de Carrera'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    puede_modificar_inscripcion(request, inscripcion)
                    data['form'] = RetiradoMateriaForm()
                    return render(request, "inscripciones/retirocarrera.html", data)
                except Exception as ex:
                    pass

            elif action == 'cambionivel':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_inscripciones')
                    data['title'] = u'Cambio de nivel malla'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    puede_modificar_inscripcion(request, inscripcion)
                    data['form'] = CambionivelmallaForm()
                    return render(request, "inscripciones/cambionivel.html", data)
                except Exception as ex:
                    pass

            elif action == 'cambiocategoria':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_inscripciones')
                    data['title'] = u'Cambio de tipo de inscripción'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    puede_modificar_inscripcion(request, inscripcion)
                    data['form'] = InscripcionTipoInscripcionForm()
                    return render(request, "inscripciones/cambiotipoinscripcion.html", data)
                except Exception as ex:
                    pass

            elif action == 'proyectos':
                try:
                    data['title'] = u'Proyecto de vinculacion'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    data['proyectos'] = inscripcion.participanteproyectovinculacion_set.all().order_by('-proyecto__fin')
                    return render(request, "inscripciones/proyectos.html", data)
                except Exception as ex:
                    pass

            elif action == 'cursos':
                try:
                    data['title'] = u'Cursos y Escuelas Complementarias'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    data['cursos'] = inscripcion.matriculacursoescuelacomplementaria_set.all()
                    return render(request, "inscripciones/cursos.html", data)
                except Exception as ex:
                    pass

            elif action == 'preinscripcion':
                try:
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    persona = inscripcion.persona
                    data['title'] = u'Pre incripciones'
                    solicitudes = ExamenComplexivo.objects.filter(inscripcion__persona=persona)
                    data['solicitudes'] = solicitudes

                    return render(request, "inscripciones/preinscripcion.html", data)
                except Exception as ex:
                    pass

            elif action == 'add':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_inscripciones')
                    data['title'] = u'Nueva inscripción'
                    form = InscripcionForm()
                    if 'id' in request.GET:
                        if PreInscrito.objects.filter(id=request.GET['id'], inscripcion=None).exists():
                            preinscrito = PreInscrito.objects.get(id=request.GET['id'])
                            data['preinscrito_id'] = preinscrito
                            form.initial = {'fecha': datetime.now().date(),
                                            'nacimiento': datetime.now().date(),
                                            'fecha_ingreso': datetime.now().date(),
                                            'nombres': preinscrito.nombres,
                                            'apellido1': preinscrito.apellido1,
                                            'apellido2': preinscrito.apellido2,
                                            'cedula': preinscrito.cedula,
                                            'sexo': preinscrito.sexo,
                                            'telefono': preinscrito.telefono_celular,
                                            'telefono_conv': preinscrito.telefono_domicilio,
                                            'telefono_trabajo': preinscrito.telefono_trabajo,
                                            'prenivelacion': False if preinscrito.pre else True,
                                            'email': preinscrito.email,
                                            'observacionespre': preinscrito.institucion.nombre if preinscrito.institucion else '',
                                            # 'colegio': preinscrito.institucion.nombre if preinscrito.institucion else '',
                                            'carrera': preinscrito.carrera,
                                            'direccion': preinscrito.direccion,
                                            'comoseinformo': preinscrito.comoseinformo,
                                            'comoseinformootras': preinscrito.comoseinformootro}
                    form.nuevo()
                    miscarreras = persona.mis_carreras()
                    if UTILIZA_GRUPOS_ALUMNOS:
                        form.con_grupos(Grupo.objects.filter(carrera__in=miscarreras))
                    else:
                        form.sin_grupos(miscarreras)
                    if EMAIL_INSTITUCIONAL_AUTOMATICO:
                        form.emailautomatico()
                    data['form'] = form
                    data['email_institucional_automatico'] = EMAIL_INSTITUCIONAL_AUTOMATICO
                    data['utiliza_grupos_alumnos'] = UTILIZA_GRUPOS_ALUMNOS
                    data['dominio'] = EMAIL_DOMAIN
                    data['preguntas_inscripcion'] = PREGUNTAS_INSCRIPCION
                    data['correo_obligatorio'] = CORREO_OBLIGATORIO
                    return render(request, "inscripciones/add.html", data)
                except Exception as ex:
                    pass

            elif action == 'adicionarotracarrera':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_inscripciones')
                    data['title'] = u'Inscripción de alumno en otra carrera'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    documentos = inscripcion.documentos_entregados()
                    tesdrive = inscripcion.documentos_tesdrive()
                    form = NuevaInscripcionForm(initial={'fecha': datetime.now().date(),
                                                         'prenivelacion': documentos.pre,
                                                         # 'copiarecord': inscripcion.recordacademico_set.all(),
                                                         'observacionespre': documentos.observaciones_pre,
                                                         'titulo': documentos.titulo,
                                                         'acta': documentos.acta,
                                                         'cedula2': documentos.cedula,
                                                         'votacion': documentos.votacion,
                                                         'actaconv': documentos.actaconv,
                                                         'partida_nac': documentos.partida_nac,
                                                         'fotos': documentos.fotos,
                                                         'licencia': tesdrive.licencia,
                                                         'record': tesdrive.record,
                                                         'certificado_tipo_sangre': tesdrive.certificado_tipo_sangre,
                                                         'prueba_psicosensometrica': tesdrive.prueba_psicosensometrica,
                                                         'certificado_estudios': tesdrive.certificado_estudios})
                    miscarreras = persona.mis_carreras()
                    data['form'] = form
                    data['dominio'] = EMAIL_DOMAIN
                    return render(request, "inscripciones/adicionarotracarrera.html", data)
                except Exception as ex:
                    pass

            elif action == 'edit':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_inscripciones')
                    data['title'] = u'Editar inscripción'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    # puede_modificar_inscripcion(request, inscripcion)
                    documentos = inscripcion.documentos_entregados()
                    tesdrive = inscripcion.documentos_tesdrive()
                    preguntas_inscripcion = inscripcion.preguntas_inscripcion()
                    perfil = inscripcion.persona.mi_perfil()
                    if inscripcion.unidadeducativa_id is None:
                        data['unid'] = 0
                    else:
                        data['unid'] = inscripcion.unidadeducativa_id
                    form = InscripcionForm(initial={'nombres': inscripcion.persona.nombres,
                                                    'apellido1': inscripcion.persona.apellido1,
                                                    'apellido2': inscripcion.persona.apellido2,
                                                    'cedula': inscripcion.persona.cedula,
                                                    'pasaporte': inscripcion.persona.pasaporte,
                                                    'centroinformacion': inscripcion.centroinformacion,
                                                    'paisnacimiento': inscripcion.persona.paisnacimiento,
                                                    'provincianacimiento': inscripcion.persona.provincianacimiento,
                                                    'cantonnacimiento': inscripcion.persona.cantonnacimiento,
                                                    'parroquianacimiento': inscripcion.persona.parroquianacimiento,
                                                    'nacionalidad': inscripcion.persona.nacionalidad,
                                                    'nacimiento': inscripcion.persona.nacimiento,
                                                    'sexo': inscripcion.persona.sexo,
                                                    'lgtbi': inscripcion.persona.lgtbi,
                                                    'raza': perfil.raza_id,
                                                    'nacionalidadindigena': perfil.nacionalidadindigena_id,
                                                    'sangre': inscripcion.persona.sangre,
                                                    'pais': inscripcion.persona.pais,
                                                    'provincia': inscripcion.persona.provincia,
                                                    'canton': inscripcion.persona.canton,
                                                    'parroquia': inscripcion.persona.parroquia,
                                                    'sector': inscripcion.persona.sector,
                                                    'direccion': inscripcion.persona.direccion,
                                                    'direccion2': inscripcion.persona.direccion2,
                                                    'num_direccion': inscripcion.persona.num_direccion,
                                                    'telefono': inscripcion.persona.telefono,
                                                    'telefono_conv': inscripcion.persona.telefono_conv,
                                                    'email': inscripcion.persona.email.strip(),
                                                    'emailinst': inscripcion.persona.emailinst.strip(),
                                                    'fecha': inscripcion.fecha,
                                                    'grupo': inscripcion.grupo(),
                                                    'sede': inscripcion.sede,
                                                    'carrera': inscripcion.carrera,
                                                    'modalidad': inscripcion.modalidad,
                                                    'sesion': inscripcion.sesion,
                                                    # 'colegio': inscripcion.colegio,
                                                    # 'unidadeducativa': inscripcion.unidadeducativa,
                                                    'especialidad': inscripcion.especialidad,
                                                    'identificador': inscripcion.identificador,
                                                    'prenivelacion': documentos.pre,
                                                    'observacionespre': documentos.observaciones_pre,
                                                    'comoseinformo': preguntas_inscripcion.comoseinformo,
                                                    'comoseinformootras': preguntas_inscripcion.comoseinformootras,
                                                    'razonesmotivaron': preguntas_inscripcion.razonesmotivaron,
                                                    'titulo': documentos.titulo,
                                                    'acta': documentos.acta,
                                                    'certificado_estudios': tesdrive.certificado_estudios,
                                                    'cedula2': documentos.cedula,
                                                    'votacion': documentos.votacion,
                                                    'partida_nac': documentos.partida_nac,
                                                    'actaconv': documentos.actaconv,
                                                    'fotos': documentos.fotos,
                                                    'licencia': tesdrive.licencia,
                                                    'record': tesdrive.record,
                                                    'certificado_tipo_sangre': tesdrive.certificado_tipo_sangre,
                                                    'prueba_psicosensometrica': tesdrive.prueba_psicosensometrica,
                                                    'itinerario': inscripcion.itinerario,
                                                    # 'ppl': inscripcion.persona.ppl,'observacionppl': inscripcion.persona.observacionppl
                                                    })
                    form.editar(inscripcion)
                    form.sin_trabajo()
                    if not puede_realizar_accion_afirmativo(request, 'sga.puede_modificar_itinerario'):
                        form.sin_itinerario()
                    miscarreras = persona.mis_carreras()
                    if UTILIZA_GRUPOS_ALUMNOS:
                        form.con_grupos(Grupo.objects.filter(carrera__in=miscarreras))
                    else:
                        form.sin_grupos(miscarreras)
                    # data['idppl'] = 1 if inscripcion.persona.ppl else 0
                    data['utiliza_grupos_alumnos'] = UTILIZA_GRUPOS_ALUMNOS
                    data['form'] = form
                    data['email_domain'] = EMAIL_DOMAIN
                    data['email_institucional_automatico'] = EMAIL_INSTITUCIONAL_AUTOMATICO
                    data['preguntas_inscripcion'] = PREGUNTAS_INSCRIPCION
                    data['correo_obligatorio'] = CORREO_OBLIGATORIO
                    return render(request, "inscripciones/edit.html", data)
                except Exception as ex:
                    pass

            elif action == 'extracurricular':
                try:
                    data['title'] = u'Actividades extracurriculares'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    data['pasantias'] = inscripcion.pasantias()
                    data['talleres'] = inscripcion.talleres()
                    data['practicas'] = inscripcion.practicas()
                    data['vccs'] = inscripcion.vcc()
                    return render(request, "inscripciones/extracurricular.html", data)
                except Exception as ex:
                    pass

            elif action == 'record':
                try:
                    data['title'] = u'Registro académico'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    data['records'] = inscripcion.recordacademico_set.filter(status=True).order_by('asignaturamalla__nivelmalla', 'asignatura', 'fecha')
                    data['total_creditos'] = inscripcion.total_creditos()
                    data['total_creditos_malla'] = inscripcion.total_creditos_malla()
                    data['total_creditos_modulos'] = inscripcion.total_creditos_modulos()
                    data['total_creditos_otros'] = inscripcion.total_creditos_otros()
                    data['total_horas'] = inscripcion.total_horas()
                    data['promedio'] = inscripcion.promedio_record()
                    data['aprobadas'] = inscripcion.recordacademico_set.filter(aprobada=True, valida=True, status=True).count()
                    data['reprobadas'] = inscripcion.recordacademico_set.filter(aprobada=False, valida=True, status=True).count()
                    data['reporte_0'] = obtener_reporte("record_alumno")
                    data['idasignaturasingles'] = []
                    data['add_record'] = variable_valor('ADICIONAR_RECORD')
                    data['edit_record'] = variable_valor('EDITAR_RECORD')

                    # if request.user.has_perm('sga.puede_modificar_ingles'):
                    data['idasignaturasingles'] = AsignaturaMalla.objects.values_list('asignatura_id', flat=True).filter(status=True, malla__carrera__id=34).distinct()
                    # if request.user.has_perm('sga.puede_modificar_computacion'):
                    #     data['idasignaturas'] = AsignaturaMalla.objects.values_list('asignatura_id', flat=True).filter(status=True, malla__carrera__id=37).distinct()
                    # data['idasignaturasno'] = AsignaturaMalla.objects.values_list('asignatura_id', flat=True).filter(status=True, malla__carrera__id__in=[34,37]).distinct()
                    data['tiene_permiso'] = persona.persona_tiene_permiso(inscripcion.id)
                    return render(request, "inscripciones/record.html", data)
                except Exception as ex:
                    return HttpResponseRedirect(f"/inscripciones?id={request.GET['id']}&info={ex.__str__()}")

            elif action == 'actvidadescomplementarias':
                try:
                    data['title'] = u'Registro academico'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    data['inscrito'] = actividadesinscritas = PaeInscripcionActividades.objects.select_related().filter(matricula__inscripcion=inscripcion, status=True, matricula__status=True).order_by('matricula__nivel__periodo__id')
                    return render(request, "inscripciones/actividadescomplementarias.html", data)
                except Exception as ex:
                    pass

            elif action == 'historico':
                try:
                    data['title'] = u'Historico de notas'
                    data['inscripcion'] = Inscripcion.objects.get(pk=request.GET['id'])
                    data['record'] = record = RecordAcademico.objects.get(pk=request.GET['rec'])
                    data['historicos'] = record.historicorecordacademico_set.all().order_by('-fecha')
                    data['edit_record'] = variable_valor('EDITAR_RECORD')
                    return render(request, "inscripciones/historico.html", data)
                except Exception as ex:
                    pass

            elif action == 'editfechamateria':
                try:
                    data['title'] = u'Editar fecha de la materia'
                    data['record'] = record = RecordAcademico.objects.get(pk=request.GET['rec'])
                    # data['profesor'] = Profesor.objects.get(pk=request.GET['pofesorid'])
                    data['form'] = FechaMateriaRecordForm(initial={'inicio': record.fechainicio,
                                                                   'fin': record.fechafin})
                    return render(request, "inscripciones/editfechamateriarecord.html", data)
                except Exception as ex:
                    pass

            elif action == 'addrecord':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_records')
                    data['title'] = u'Adicionar registro academico'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    puede_modificar_inscripcion(request, inscripcion)
                    if personasesion.usuario.id in [43139, 43138, 46006 ]:
                        form = RecordAcademicoForm(initial={'asistencia':'100', 'observaciones':'MIGRACIÓN FASE II'})
                    else:
                        form = RecordAcademicoForm()
                    form.solo_mallas_carrera(inscripcion,personasesion,request)
                    data['form'] = form
                    return render(request, "inscripciones/addrecord.html", data)
                except Exception as ex:
                    pass

            elif action == 'addrecordhomologada':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_homologaciones')
                    data['title'] = u'Adicionar homologación'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    puede_modificar_inscripcion(request, inscripcion)
                    form = RecordAcademicoForm()
                    form.homologacion()
                    data['form'] = form
                    return render(request, "inscripciones/addrecordhomologada.html", data)
                except Exception as ex:
                    pass

            elif action == 'addhistorico':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_records')
                    data['title'] = u'Adicionar historico de registro academico'
                    data['record'] = record = RecordAcademico.objects.get(pk=request.GET['idr'])
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    puede_modificar_inscripcion(request, inscripcion)
                    form = HistoricoRecordAcademicoForm(initial={"fecha": datetime.now().date(),
                                                                 "creditos": 0,
                                                                 "horas": 0,
                                                                 "nota": 0,
                                                                 "asistencia": 0})
                    form.solo_asignatura(record.asignatura.id)
                    data['form'] = form
                    return render(request, "inscripciones/addhistorico.html", data)
                except Exception as ex:
                    pass

            elif action == 'matricular':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_matriculas')
                    inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    puede_modificar_inscripcion(request, inscripcion)
                    periodo = request.session['periodo']
                    if MATRICULACION_LIBRE:
                        hoy = datetime.now().date()
                        #se le quito , periodo__tipo__id=TIPO_PERIODO_REGULAR en el if
                        nivelaux = Nivel.objects.filter(nivellibrecoordinacion__coordinacion__carrera=inscripcion.carrera,
                                                        nivellibrecoordinacion__coordinacion__sede=inscripcion.sede, modalidad=inscripcion.modalidad, cerrado=False,
                                                        fin__gte=hoy, periodo=periodo)
                        if nivelaux.exists():
                            nivel = Nivel.objects.filter(nivellibrecoordinacion__coordinacion__carrera=inscripcion.carrera, nivellibrecoordinacion__coordinacion__sede=inscripcion.sede, modalidad=inscripcion.modalidad, cerrado=False, fin__gte=hoy, periodo=periodo).order_by('-fin')[0]
                            return HttpResponseRedirect('/matriculas?action=addmatriculalibre&id=' + str(nivel.id) + '&iid=' + str(inscripcion.id))
                    return HttpResponseRedirect("/inscripciones?id=" + request.GET['id'] + "&menserror=1")
                except Exception as ex:
                    pass

            elif action == 'edithistorico':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_records')
                    data['title'] = u'Editar historico de registro academico'
                    historico = HistoricoRecordAcademico.objects.get(pk=request.GET['id'])
                    # puede_modificar_inscripcion(request, historico.inscripcion)
                    form = HistoricoRecordAcademicoForm(initial={"asignatura": historico.asignatura,
                                                                 "creditos": historico.creditos,
                                                                 "horas": historico.horas,
                                                                 "nota": historico.nota,
                                                                 "asistencia": historico.asistencia,
                                                                 "fecha": historico.fecha,
                                                                 "noaplica": historico.noaplica,
                                                                 "aprobada": historico.aprobada,
                                                                 "valida": historico.valida,
                                                                 "validapromedio": historico.validapromedio,
                                                                 "convalidacion": historico.convalidacion,
                                                                 "homologada": historico.homologada,
                                                                 "observaciones": historico.observaciones,
                                                                 "suficiencia": historico.suficiencia})
                    puede_editar = variable_valor('ADICIONAR_RECORD')
                    form.editar(historico, puede_editar)
                    data['form'] = form
                    data['historico'] = historico
                    return render(request, "inscripciones/edithistorico.html", data)
                except Exception as ex:
                    pass

            elif action == 'delrecord':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_records')
                    data['title'] = u'Eliminar registro academico'
                    data['record'] = record = RecordAcademico.objects.get(pk=request.GET['id'])
                    puede_modificar_inscripcion(request, record.inscripcion)
                    return render(request, "inscripciones/delrecord.html", data)
                except Exception as ex:
                    pass

            elif action == 'delhistorico':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_records')
                    data['title'] = u'Eliminar historico de registro academico'
                    data['historico'] = historico = HistoricoRecordAcademico.objects.get(pk=request.GET['id'])
                    puede_modificar_inscripcion(request, historico.inscripcion)
                    return render(request, "inscripciones/delhistorico.html", data)
                except Exception as ex:
                    pass

            elif action == 'convalidar':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_homologaciones')
                    data['title'] = u'Homologación de materia'
                    record = RecordAcademico.objects.get(pk=request.GET['id'])
                    puede_modificar_inscripcion(request, record.inscripcion)
                    historico = record.mi_historico()
                    convalidacion = historico.datos_convalidacion()
                    data['form'] = ConvalidacionInscripcionForm(initial={'centro': convalidacion.centro,
                                                                         'carrera': convalidacion.carrera,
                                                                         'asignatura': convalidacion.asignatura,
                                                                         'anno': convalidacion.anno,
                                                                         'nota_ant': convalidacion.nota_ant,
                                                                         'nota_act': convalidacion.nota_act,
                                                                         'creditos': convalidacion.creditos,
                                                                         'observaciones': convalidacion.observaciones})
                    data['convalidacion'] = convalidacion
                    return render(request, "inscripciones/convalidar.html", data)
                except Exception as ex:
                    transaction.set_rollback(True)
                    pass

            elif action == 'homologar':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_homologaciones')
                    data['title'] = u'Homologacion de materia'
                    record = RecordAcademico.objects.get(pk=request.GET['id'])
                    puede_modificar_inscripcion(request, record.inscripcion)
                    historico = record.mi_historico()
                    homologacion = historico.datos_homologacion()
                    data['form'] = HomologacionInscripcionForm(initial={'carrera': homologacion.carrera,
                                                                        'asignatura': homologacion.asignatura,
                                                                        'fecha': homologacion.fecha,
                                                                        'nota_ant': homologacion.nota_ant,
                                                                        'creditos': homologacion.creditos,
                                                                        'modalidad': homologacion.modalidad,
                                                                        'observaciones': homologacion.observaciones})
                    data['record'] = record
                    data['homologacion'] = homologacion
                    return render(request, "inscripciones/homologar.html", data)
                except Exception as ex:
                    transaction.set_rollback(True)
                    pass

            elif action == 'cargarfoto':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_inscripciones')
                    data['title'] = u'Subir foto'
                    data['form'] = CargarFotoForm()
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    puede_modificar_inscripcion(request, inscripcion)
                    return render(request, "inscripciones/cargarfoto.html", data)
                except Exception as ex:
                    pass

            elif action == 'documentos':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_inscripciones')
                    data['title'] = u'Documentos y archivos'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    puede_modificar_inscripcion(request, inscripcion)
                    data['archivos'] = inscripcion.archivo_set.all().order_by('fecha')
                    data['docperfil'] = inscripcion.persona.mi_perfil()
                    data['docreligion'] = PersonaReligion.objects.filter(status=True, persona= inscripcion.persona)
                    data['docestudiante'] = inscripcion.persona
                    data['docpersonal'] = PersonaDocumentoPersonal.objects.filter(status=True, persona=inscripcion.persona).first()
                    return render(request, "inscripciones/documentos.html", data)
                except Exception as ex:
                    pass

            elif action == 'adddocumento':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_inscripciones')
                    data['title'] = u'Adicionar archivos'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    puede_modificar_inscripcion(request, inscripcion)
                    data['form'] = DocumentoInscripcionForm()
                    return render(request, "inscripciones/adddocumento.html", data)
                except Exception as ex:
                    pass

            elif action == 'importar':
                try:
                    puede_realizar_accion(request, 'sga.puede_importar_inscripciones')
                    data['title'] = u'Importar inscripciones'
                    data['form'] = ImportarArchivoXLSForm()
                    return render(request, "inscripciones/importar.html", data)
                except Exception as ex:
                    pass

            elif action == 'deldocumento':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_inscripciones')
                    data['title'] = u'Eliminar archivo o documento'
                    data['archivo'] = archivo = Archivo.objects.get(pk=request.GET['id'])
                    puede_modificar_inscripcion(request, archivo.inscripcion)
                    return render(request, "inscripciones/deldocumento.html", data)
                except Exception as ex:
                    pass

            elif action == 'fechainicioconvalidacion':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_inscripciones')
                    data['title'] = u'Fecha inicio convalidacion'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    puede_modificar_inscripcion(request, inscripcion)
                    data['form'] = FechaInicioConvalidacionInscripcionForm(initial={'fecha': inscripcion.fechainicioconvalidacion if inscripcion.fechainicioconvalidacion else datetime.now().date()})
                    return render(request, "inscripciones/fechainicioconvalidacion.html", data)
                except Exception as ex:
                    pass

            elif action == 'fechainicioprimernivel':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_inscripciones')
                    data['title'] = u'Fecha inicio primer nivel'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    puede_modificar_inscripcion(request, inscripcion)
                    data['form'] = FechaInicioPrimerNivelInscripcionForm(initial={'fecha': inscripcion.fechainicioprimernivel if inscripcion.fechainicioprimernivel else datetime.now().date()})
                    return render(request, "inscripciones/fechainicioprimernivel.html", data)
                except Exception as ex:
                    pass

            elif action == 'fechainiciocarrera':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_inscripciones')
                    data['title'] = u'Fecha inicio carrera'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    puede_modificar_inscripcion(request, inscripcion)
                    data['form'] = FechaInicioCarreraInscripcionForm(initial={'fecha': inscripcion.fechainiciocarrera if inscripcion.fechainiciocarrera else datetime.now().date()})
                    return render(request, "inscripciones/fechainiciocarrera.html", data)
                except Exception as ex:
                    pass

            elif action == 'alumalla':
                try:
                    listas_asignaturasmallas = []
                    data['title'] = u'Malla del alumno'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    data['inscripcion_malla'] = inscripcionmalla = inscripcion.malla_inscripcion()
                    data['malla'] = malla = inscripcionmalla.malla
                    nivelmalla = NivelMalla.objects.all().order_by('id')
                    # data['asignaturasmallas'] = [(x, inscripcion.aprobadaasignatura(x)) for x in AsignaturaMalla.objects.filter(malla=malla)]
                    listadoasignaturamalla = AsignaturaMalla.objects.filter(malla=malla, status=True).exclude(tipomateria_id=3)
                    xyz =[1,2,3]
                    if inscripcion.itinerario and inscripcion.itinerario>0:
                        xyz.remove(inscripcion.itinerario)
                        listadoasignaturamalla = listadoasignaturamalla.exclude(itinerario__in=xyz)
                    for x in listadoasignaturamalla.exclude(ejeformativo_id=4):
                        listas_asignaturasmallas.append([x, inscripcion.aprobadaasignatura(x)])

                    for ni in nivelmalla:
                        listadooptativa = listadoasignaturamalla.filter(nivelmalla=ni, ejeformativo_id=4)
                        if inscripcion.recordacademico_set.filter(asignatura__in=listadooptativa.values_list('asignatura__id', flat=True)).exists():
                            for d in listadooptativa:
                                if inscripcion.aprobadaasignatura(d):
                                    listas_asignaturasmallas.append([d, inscripcion.aprobadaasignatura(d)])
                        else:
                            for d in listadooptativa:
                                listas_asignaturasmallas.append([d, inscripcion.aprobadaasignatura(d)])
                    data['asignaturasmallas'] = listas_asignaturasmallas
                    data['nivelesdemallas'] = nivelmalla = nivelmalla.filter(pk__in=listadoasignaturamalla.values_list('nivelmalla__id', flat=True).distinct())
                    data['ejesformativos'] = EjeFormativo.objects.filter(pk__in=listadoasignaturamalla.values_list('ejeformativo__id', flat=True).distinct()).order_by('nombre')
                    resumenniveles = [{'id': x.id, 'horas': x.total_horas(malla), 'creditos': x.total_creditos(malla)} for x in nivelmalla.order_by('id')]
                    data['resumenes'] = resumenniveles
                    return render(request, "inscripciones/alumalla.html", data)
                except Exception as ex:
                    transaction.set_rollback(True)
                    pass

            elif action == 'alumallahistorico':
                try:
                    data['title'] = u'Malla del alumno'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    data['inscripcion_malla'] = inscripcionmalla = inscripcion.malla_inscripcion()
                    data['malla'] = malla = inscripcionmalla.malla
                    data['nivelesdemallas'] = NivelMalla.objects.filter(status=True).order_by('id')
                    data['ejesformativos'] = EjeFormativo.objects.filter(status=True).order_by('nombre')
                    # data['asignaturasmallas'] = [(x, inscripcion.aprobadaasignatura(x)) for x in AsignaturaMalla.objects.filter(malla=malla)]
                    # resumenniveles = [{'id': x.id, 'horas': x.total_horas(malla), 'creditos': x.total_creditos(malla)} for x in NivelMalla.objects.all().order_by('id')]
                    # data['resumenes'] = resumenniveles
                    return render(request, "inscripciones/alumallahistorico.html", data)
                except Exception as ex:
                    transaction.set_rollback(True)
                    pass

            elif action == 'novalidar':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_records')
                    data['title'] = u'No considerar créditos'
                    data['record'] = record = RecordAcademico.objects.get(pk=request.GET['id'])
                    puede_modificar_inscripcion(request, record.inscripcion)
                    data['form'] = ConsiderarForm()
                    return render(request, "inscripciones/novalidar.html", data)
                except Exception as ex:
                    pass

            elif action == 'validar':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_records')
                    data['title'] = u'Considerar créditos'
                    data['record'] = record = RecordAcademico.objects.get(pk=request.GET['id'])
                    puede_modificar_inscripcion(request, record.inscripcion)
                    data['form'] = ConsiderarForm()
                    return render(request, "inscripciones/validar.html", data)
                except Exception as ex:
                    pass

            elif action == 'novalidarpromedio':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_records')
                    data['title'] = u'No considerar para promedio'
                    data['record'] = record = RecordAcademico.objects.get(pk=request.GET['id'])
                    puede_modificar_inscripcion(request, record.inscripcion)
                    data['form'] = ConsiderarForm()
                    return render(request, "inscripciones/novalidarpromedio.html", data)
                except Exception as ex:
                    pass

            elif action == 'validarpromedio':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_records')
                    data['title'] = u'Considerar para promedio'
                    data['record'] = record = RecordAcademico.objects.get(pk=request.GET['id'])
                    puede_modificar_inscripcion(request, record.inscripcion)
                    data['form'] = ConsiderarForm()
                    return render(request, "inscripciones/validarpromedio.html", data)
                except Exception as ex:
                    pass

            elif action == 'cambiogrupo':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_inscripciones')
                    data['title'] = u'Cambio de grupo'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    puede_modificar_inscripcion(request, inscripcion)
                    form = CambioGrupoForm()
                    form.cambio_grupo(inscripcion)
                    data['form'] = form

                    return render(request, "inscripciones/cambiogrupo.html", data)
                except Exception as ex:
                    pass

            elif action == 'recalcularcreditos':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_records')
                    inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    puede_modificar_inscripcion(request, inscripcion)
                    inscripcion.actualizar_creditos()
                    return HttpResponseRedirect("/inscripciones?action=record&id=" + request.GET['id'])
                except Exception as ex:
                    transaction.set_rollback(True)
                    return HttpResponseRedirect(f"/inscripciones?id={request.GET['id']}&info={ex.__str__()}")

            elif action == 'recalcularniveles':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_records')
                    inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    puede_modificar_inscripcion(request, inscripcion)
                    inscripcion.actualizar_niveles_records()
                    inscripcion.actualizar_nivel()
                    return HttpResponseRedirect("/inscripciones?action=record&id=" + request.GET['id'])
                except Exception as ex:
                    transaction.set_rollback(True)
                    return HttpResponseRedirect(f"/inscripciones?id={request.GET['id']}&info={ex.__str__()}")

            elif action == 'desactivar':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_inscripciones')
                    data['title'] = u'Desactivar usuario'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    puede_modificar_inscripcion(request, inscripcion)
                    return render(request, "inscripciones/desactivar.html", data)
                except Exception as ex:
                    pass

            elif action == 'activar':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_inscripciones')
                    data['title'] = u'Activar usuario'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    puede_modificar_inscripcion(request, inscripcion)
                    return render(request, "inscripciones/activar.html", data)
                except Exception as ex:
                    pass

            elif action == 'desactivarperfil':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_administrativos')
                    data['title'] = u'Desactivar perfil de usuario'
                    data['inscripcion'] = Inscripcion.objects.get(pk=request.GET['id'])
                    return render(request, "inscripciones/desactivarperfil.html", data)
                except Exception as ex:
                    pass

            elif action == 'bloqueomatricula':
                try:
                    puede_realizar_accion(request, 'sga.puede_bloquear_matricula')
                    data['matricula'] = mnatricula = Matricula.objects.get(pk=request.GET['id'])
                    data['inscripcion'] = Inscripcion.objects.get(pk=mnatricula.inscripcion.id)
                    data['title'] = u'Desbloquear matricula' if mnatricula.bloqueomatricula else u'Bloquear matricula'
                    return render(request, "inscripciones/bloqueomatricula.html", data)
                except Exception as ex:
                    pass

            elif action == 'activarperfil':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_administrativos')
                    data['title'] = u'Activar perfil de usuario'
                    data['inscripcion'] = Inscripcion.objects.get(pk=request.GET['id'])
                    return render(request, "inscripciones/activarperfil.html", data)
                except Exception as ex:
                    pass

            elif action == 'resetear':
                try:
                    puede_realizar_accion(request, 'sga.puede_resetear_clave')
                    data['title'] = u'Resetear clave del usuario'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    # puede_modificar_inscripcion(request, inscripcion)
                    return render(request, "inscripciones/resetear.html", data)
                except Exception as ex:
                    pass

            elif action == 'aplicab2':
                try:
                    data['title'] = u'B2'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    data['valor'] = request.GET['valor']
                    return render(request, "inscripciones/aplicab2.html", data)
                except Exception as ex:
                    pass

            elif action == 'actividades':
                try:
                    data['title'] = u'Actividades extracurriculares'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    data['actividades'] = inscripcion.participanteactividadextracurricular_set.all().order_by('-actividad__fechafin')
                    return render(request, "inscripciones/actividades.html", data)
                except Exception as ex:
                    pass

            elif action == 'horario':
                try:
                    data['title'] = u'Horario del estudiante'
                    data['matricula'] = matricula = Matricula.objects.get(pk=request.GET['id'])
                    data['semana'] = [[1, 'Lunes'], [2, 'Martes'], [3, 'Miercoles'], [4, 'Jueves'], [5, 'Viernes'], [6, 'Sabado'], [7, 'Domingo']]
                    hoy = datetime.now().date()
                    data['misclases'] = clases = Clase.objects.filter(activo=True, materia__materiaasignada__matricula=matricula, materia__materiaasignada__retiramateria=False, materia__materiaasignada__status=True).order_by('inicio')
                    # data['misclases'] = clases = Clase.objects.filter(activo=True, fin__gte=hoy, materia__materiaasignada__matricula=matricula, materia__materiaasignada__retiramateria=False).order_by('inicio')
                    data['inscripcion'] = matricula.inscripcion
                    data['sesiones'] = Sesion.objects.filter(turno__clase__in=clases).distinct()
                    data['idperiodo'] = periodo.id
                    return render(request, "inscripciones/horario.html", data)
                except Exception as ex:
                    pass

            elif action == 'addadministrativo':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_administrativos')
                    data['title'] = u'Crear perfil de administrativo'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    return render(request, "inscripciones/addadministrativo.html", data)
                except Exception as ex:
                    pass

            elif action == 'adddocente':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_profesores')
                    data['title'] = u'Crear perfil de profesor'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    return render(request, "inscripciones/adddocente.html", data)
                except Exception as ex:
                    pass

            elif action == 'edit_titulos':
                try:
                    data['title'] = u'Titulos profesionales del Alumno'
                    inscripcion = Inscripcion.objects.get(id=int(request.GET['id']))
                    data['id'] = inscripcion.id
                    data['idp'] = inscripcion.persona.id
                    data['personatitulouniversidad'] = personatitulouniversidad = PersonaTituloUniversidad.objects.filter(persona_id=inscripcion.persona.id)
                    return render(request, "inscripciones/edit_titulos.html", data)
                except Exception as ex:
                    pass

            elif action == 'add_titulos_profesionales':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_inscripciones')
                    data['title'] = u'Adicionar Titulos profesionales al Alumno'
                    data['idp'] = int(request.GET['idp'])
                    data['id'] = int(request.GET['id'])
                    data['form'] = PersonaTituloUniversidadForm()
                    return render(request, "inscripciones/add_titulos_profesionales.html", data)
                except Exception as ex:
                    pass

            elif action == 'edit_titulos_profesionales':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_inscripciones')
                    data['title'] = u'Editar Titulos profesionales al Alumno'
                    data['id'] = int(request.GET['id'])
                    data['idp'] = int(request.GET['idp'])
                    data['idi'] = int(request.GET['idi'])
                    data['personatitulouniversidad'] = personatitulouniversidad = PersonaTituloUniversidad.objects.get(pk=request.GET['id'])
                    data['form'] = PersonaTituloUniversidadForm(initial={'codigoregistro':personatitulouniversidad.codigoregistro,
                                                                         'fecharegresado':personatitulouniversidad.fecharegresado,
                                                                         'fecharegistro':personatitulouniversidad.fecharegistro,
                                                                         'fechaacta':personatitulouniversidad.fechaacta,
                                                                         'fechainicio':personatitulouniversidad.fechainicio,
                                                                         'tiponivel':personatitulouniversidad.tiponivel,
                                                                         'tipouniversidad':personatitulouniversidad.tipouniversidad,
                                                                         'universidad':personatitulouniversidad.universidad,
                                                                         'nombrecarrera':personatitulouniversidad.nombrecarrera})
                    return render(request, "inscripciones/edit_titulos_profesionales.html", data)
                except Exception as ex:
                    pass

            elif action == 'delete_titulos_profesionales':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_inscripciones')
                    data['title'] = u'Eliminar Titulos profesionales al Alumno'
                    data['idp'] = int(request.GET['idp'])
                    data['personatitulouniversidad'] = PersonaTituloUniversidad.objects.get(pk=request.GET['id'])
                    return render(request, "inscripciones/delete_titulos_profesionales.html", data)
                except Exception as ex:
                    pass

            elif action == 'editmallahistorica':
                try:
                    data['title'] = u'Añadir malla historica'
                    data['record'] =record= RecordAcademico.objects.get(pk=request.GET['rec'])
                    form = MallaHistoricaForm(initial={"asignaturamallahistorico": record.asignaturamallahistorico,})
                    form.solo_mallas_carrera(record.inscripcion)
                    data['form'] = form
                    return render(request, "inscripciones/editmallahistorica.html", data)
                except Exception as ex:
                    pass

            elif action == 'cambiarbloqueomatricula':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_inscripciones')
                    data['title'] = u'Cambiar estado matricula'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    return render(request, "inscripciones/cambiarbloqueomatricula.html", data)
                except Exception as ex:
                    pass

            elif action == 'desactivarperfilusuario':
                try:
                    puede_realizar_accion(request, 'sga.puede_aprobar_perfil_usuario')
                    data['title'] = u'Desactivar Perfil Usuario'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    return render(request, "inscripciones/desactivarperfilusuario.html", data)
                except Exception as ex:
                    pass

            elif action == 'activarperfilusuario':
                try:
                    puede_realizar_accion(request, 'sga.puede_aprobar_perfil_usuario')
                    data['title'] = u'Activar Perfil Usuario'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    return render(request, "inscripciones/activarperfilusuario.html", data)
                except Exception as ex:
                    pass

            elif action == 'soportes':
                try:
                    data['title'] = u'Listado Soportes'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    data['soportes'] = VirtualSoporteUsuario.objects.filter(status=True)
                    return render(request, "inscripciones/soportes.html", data)
                except Exception as ex:
                    pass

            elif action == 'resetear_clave_pregrado_virtual':
                try:
                    data['title'] = u'Resetear clave del usuario'
                    data['inscripcion'] = Inscripcion.objects.get(pk=request.GET['id'])
                    return render(request, "inscripciones/resetear_clave_pregrado_virtual.html", data)
                except Exception as ex:
                    pass

            elif action == 'reportezurdos':
                try:
                    __author__ = 'Unemi'
                    title = easyxf( 'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('hoja1')
                    ws.write_merge(0, 0, 0, 7, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=reporte_personas_zurdas' + random.randint(1, 10000).__str__() + '.xls'
                    columns = [
                        (u"CEDULA", 3000),
                        (u"NOMBRES", 6000),
                        (u"CARRERA", 6000),
                        (u"MODALIDAD", 6000),
                        (u"JORNADA", 6000),
                        (u"PARALELO", 6000),
                    ]
                    row_num = 3
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    cursor = connections['sga_select'].cursor()
                    idp = periodo.id.__str__()
                    sql = """
                        SELECT DISTINCT p.cedula, p.apellido1||' '|| p.apellido2||' '|| p.nombres AS nombres, 
                        c.nombre AS carrera, 
                        (CASE WHEN c.modalidad=1 THEN 'PRESENCIAL' WHEN c.modalidad=2 THEN 'SEMI PRESENCIAL' WHEN c.modalidad=3 THEN 'EN LÍNEA' ELSE 'PRESENCIAL' END) AS modalidad
                        , ses.nombre AS jornada , para.nombre AS paralelo
                        FROM sga_persona p
                        INNER JOIN sga_inscripcion i ON i.persona_id=p.id
                        INNER JOIN sga_matricula matr ON matr.inscripcion_id=i.id
                        INNER JOIN sga_nivel niv ON niv.id=matr.nivel_id
                        INNER JOIN sga_sesion ses ON ses.id=niv.sesion_id
                        INNER JOIN sga_paralelo para ON para.id=matr.paralelo_id
                        INNER JOIN sga_carrera c ON c.id=i.carrera_id
                        INNER JOIN sga_coordinacion_carrera cc ON cc.carrera_id=c.id
                        where p.eszurdo=TRUE AND matr.cerrada=FALSE AND niv.periodo_id=%s
                    """ % (idp)
                    cursor.execute(sql)
                    results = cursor.fetchall()
                    row_num = 4
                    for r in results:
                        i = 0
                        campo1 = r[0]
                        campo2 = r[1]
                        campo3 = r[2]
                        campo4 = r[3]
                        campo5 = r[4]
                        campo6 = r[5]
                        ws.write(row_num, 0, campo1.__str__(), font_style2)
                        ws.write(row_num, 1, campo2.__str__(), font_style2)
                        ws.write(row_num, 2, campo3.__str__(), font_style2)
                        ws.write(row_num, 3, campo4.__str__(), font_style2)
                        ws.write(row_num, 4, campo5.__str__(), font_style2)
                        ws.write(row_num, 5, campo6.__str__(), font_style2)
                        row_num += 1
                    wb.save(response)
                    return response
                except Exception as ex:
                    pass

            elif action == 'reporte_estudianteshijos':
                try:
                    __author__ = 'Unemi'
                    title = easyxf(
                        'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('hoja1')
                    ws.write_merge(0, 0, 0, 7, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=reporte_personas_con_hijos' + random.randint(1,
                                                                                                                         10000).__str__() + '.xls'
                    columns = [
                        (u"CEDULA", 3000),
                        (u"NOMBRES", 6000),
                        (u"CARRERA", 6000),
                        (u"MODALIDAD", 6000),
                        (u"JORNADA", 6000),
                        (u"PARALELO", 6000),
                        (u'PARENTESCO', 6000),
                        (u'RANGO EDAD', 6000),
                        (u'NOMBRE HIJO/A', 6000)
                    ]
                    row_num = 3
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    cursor = connections['sga_select'].cursor()
                    idp = periodo.id.__str__()
                    sql = """SELECT DISTINCT p.cedula, p.apellido1||' '|| p.apellido2||' '|| p.nombres AS nombres, 
                                       c.nombre AS carrera, 
                                       (CASE WHEN c.modalidad=1 THEN 'PRESENCIAL' WHEN c.modalidad=2 THEN 'SEMI PRESENCIAL' WHEN c.modalidad=3 THEN 'EN LÍNEA' ELSE 'PRESENCIAL' END) AS modalidad                                    
                                       , ses.nombre AS jornada , para.nombre AS paralelo,
                                       (CASE WHEN dp.parentesco_id = 11 THEN 'HIJO' WHEN dp.parentesco_id = 14 THEN 'HIJA' END) AS Parentesco,
													 dp.rangoedad AS rangoedad, dp.nombre AS NombreHijo
                                       FROM sga_persona p
                                       INNER JOIN sga_inscripcion i ON i.persona_id=p.id
                                       INNER JOIN sga_matricula matr ON matr.inscripcion_id=i.id
                                       INNER JOIN sga_nivel niv ON niv.id=matr.nivel_id
                                       INNER JOIN sga_sesion ses ON ses.id=niv.sesion_id
                                       INNER JOIN sga_paralelo para ON para.id=matr.paralelo_id
                                       INNER JOIN sga_carrera c ON c.id=i.carrera_id
                                       INNER JOIN sga_coordinacion_carrera cc ON cc.carrera_id=c.id
                                       INNER JOIN sga_personadatosfamiliares dp ON dp.persona_id = p.id
                                       where p.numeromiembrosfamilia >=1 AND (dp.parentesco_id = 11 OR dp.parentesco_id = 14) 
														AND matr.cerrada=FALSE AND niv.periodo_id=%s
														AND extract(year from dp.nacimiento) >= 2015
                                       GROUP BY
                                       	dp.parentesco_id,
                                       	p.cedula, p.apellido1, p.apellido2, p.nombres, 
                                       	c.nombre, c.modalidad, ses.nombre, para.nombre,
                                       	p.nacimiento, dp.rangoedad, dp.nacimiento, dp.nombre,
                                       	p.numeromiembrosfamilia
                                       """ % (idp)
                    cursor.execute(sql)
                    results = cursor.fetchall()
                    row_num = 4
                    for r in results:
                        i = 0
                        campo1 = r[0]
                        campo2 = r[1]
                        campo3 = r[2]
                        campo4 = r[3]
                        campo5 = r[4]
                        campo6 = r[5]
                        campo7 = r[6]
                        campo8 = r[7]
                        campo9 = r[8]
                        ws.write(row_num, 0, campo1.__str__(), font_style2)
                        ws.write(row_num, 1, campo2.__str__(), font_style2)
                        ws.write(row_num, 2, campo3.__str__(), font_style2)
                        ws.write(row_num, 3, campo4.__str__(), font_style2)
                        ws.write(row_num, 4, campo5.__str__(), font_style2)
                        ws.write(row_num, 5, campo6.__str__(), font_style2)
                        ws.write(row_num, 6, campo7.__str__(), font_style2)
                        ws.write(row_num, 7, campo8.__str__(), font_style2)
                        ws.write(row_num, 8, campo9.__str__(), font_style2)
                        row_num += 1
                    wb.save(response)
                    return response
                except Exception as ex:
                    pass

            elif action == 'reporte_encuestacoronavirus':
                try:
                    __author__ = 'Unemi'
                    title = easyxf( 'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('exp_xls_post_part')
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=Encuestados ' + random.randint(1, 10000).__str__() + '.xls'
                    columns = [
                        (u"Alumno", 6000),
                        (u"Cedula", 6000),
                        (u"Carrera", 10000),
                        (u"modalidad", 10000),
                        (u"Email institucional", 6000),
                        (u"Email personal", 6000),
                        (u"Telefono", 6000),
                        (u"Provincia", 6000),
                        (u"Canton", 6000),
                        (u"Parroquia", 6000),
                        (u"sector ", 6000),
                        (u"direccion 1 ", 6000),
                        (u"direccion2", 6000),
                    ]
                    row_num = 0
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    row_num = 0
                    col_num = len(columns)
                    for p in PreguntaEncuestaTecnologica.objects.filter(status=True):
                        ws.write(row_num, col_num, '%s' % p.descripcion, font_style2)
                        col_num = col_num + 1
                    row_num = 1
                    maatr=0
                    idmat=EncuestaTecnologica.objects.values_list('matricula_id').filter(status=True).distinct()
                    for matricula in Matricula.objects.filter(status=True,id__in=idmat, nivel__periodo_id=90).order_by('inscripcion__persona__apellido1', 'inscripcion__persona__apellido2','inscripcion__persona__nombres'):
                        if matricula.id!=maatr:
                            ws.write(row_num, 0, '%s' %matricula.inscripcion.persona.nombre_completo_inverso() , font_style2)
                            ws.write(row_num, 1, '%s' %matricula.inscripcion.persona.identificacion() , font_style2)
                            ws.write(row_num, 2, '%s' %matricula.inscripcion.carrera , font_style2)
                            ws.write(row_num, 3, '%s' %matricula.inscripcion.carrera.get_modalidad_display() if matricula.inscripcion.carrera.modalidad else "" , font_style2)
                            ws.write(row_num, 4, '%s' %matricula.inscripcion.persona.emailinst if matricula.inscripcion.persona.emailinst else '', font_style2)
                            ws.write(row_num, 5, '%s' %matricula.inscripcion.persona.email if matricula.inscripcion.persona.email else '', font_style2)
                            ws.write(row_num, 6, '%s' %matricula.inscripcion.persona.telefono if matricula.inscripcion.persona.telefono else '',font_style2)
                            ws.write(row_num, 7, '%s' %matricula.inscripcion.persona.provincia if matricula.inscripcion.persona.provincia else '',font_style2)
                            ws.write(row_num, 8, '%s' %matricula.inscripcion.persona.canton if matricula.inscripcion.persona.canton else '',font_style2)
                            ws.write(row_num, 9, '%s' %matricula.inscripcion.persona.parroquia if matricula.inscripcion.persona.parroquia else '',font_style2)
                            ws.write(row_num, 10, '%s' %matricula.inscripcion.persona.sector if matricula.inscripcion.persona.sector else '',font_style2)
                            ws.write(row_num, 11, '%s' %matricula.inscripcion.persona.direccion if matricula.inscripcion.persona.direccion else '',font_style2)
                            ws.write(row_num, 12, '%s' %matricula.inscripcion.persona.direccion2 if matricula.inscripcion.persona.direccion2 else '',font_style2)
                            col_num = 12 + 1
                            for r in PreguntaEncuestaTecnologica.objects.filter(status=True):
                                resu="NINGUNA"
                                if EncuestaTecnologica.objects.filter(status=True, matricula=matricula, pregunta=r).exists():
                                    resul = EncuestaTecnologica.objects.filter(status=True, matricula=matricula, pregunta=r)[0]
                                    if r.tipo==1:
                                        resu=resul.get_respuestasino_display() if resul.respuestasino else "NINGUNA"
                                    elif r.tipo ==2:
                                        resu = resul.get_respuestarango_display() if resul.respuestarango else "NINGUNA"
                                    elif r.tipo == 3:
                                        resu = resul.descripcion if resul.descripcion else "NINGUNA"
                                ws.write(row_num, col_num, '%s' % resu, font_style2)
                                col_num = col_num + 1
                            row_num += 1
                        maatr = matricula.id
                    wb.save(response)
                    return response
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos. %s"%ex})

            elif action == 'editusuario':
                try:
                    puede_realizar_accion(request, 'sga.puede_cambiar_nombre_usuario')
                    data['title'] = u'Editar nombre de usuario'
                    data['inscripcion'] =inscripcion= Inscripcion.objects.get(pk=int(encrypt(request.GET['id'])))
                    form = NombreUsuarioForm(initial={"nombre": inscripcion.persona.usuario.username,})
                    data['form'] = form
                    listausuarios=[]
                    for cont in range(10):
                        variant=cont+1
                        alfabeto = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n', 'o', 'p', 'q',
                                    'r', 's', 't', 'u', 'v', 'w', 'x', 'y', 'z', '0', '1', '2', '3', '4', '5', '6', '7',
                                    '8', '9']
                        personaestudiante=inscripcion.persona
                        s = personaestudiante.nombres.lower().split(' ')
                        while '' in s:
                            s.remove('')
                        if personaestudiante.apellido2:
                            usernamevariant = s[0][0] + personaestudiante.apellido1.lower() + personaestudiante.apellido2.lower()[0]
                        else:
                            usernamevariant = s[0][0] + personaestudiante.apellido1.lower()
                        usernamevariant = usernamevariant.replace(' ', '').replace(u'ñ', 'n').replace(u'á', 'a').replace(
                            u'é', 'e').replace(u'í', 'i').replace(u'ó', 'o').replace(u'ú', 'u')
                        usernamevariantfinal = ''
                        for letra in usernamevariant:
                            if letra in alfabeto:
                                usernamevariantfinal += letra

                        if variant > 0:
                            usernamevariantfinal += str(variant)
                        if not User.objects.values('id').filter(username=usernamevariantfinal).exists():
                            listausuarios.append(usernamevariantfinal)
                            if listausuarios.__len__() == 5:
                                break

                    data['listausuarios']=listausuarios
                    return render(request, "inscripciones/editusuario.html", data)
                except Exception as ex:
                    pass

            elif action == 'ppl':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_inscripciones')
                    data['title'] = u'Historial PPL'
                    search = None
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=int(encrypt(request.GET['id'])))
                    historialppl = HistorialPersonaPPL.objects.filter(persona=inscripcion.persona, status=True).order_by('fechaingreso')
                    paging = MiPaginador(historialppl, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['historialppl'] = page.object_list
                    return render(request, "inscripciones/historialppl.html", data)
                except Exception as ex:
                    return HttpResponseRedirect(f'{request.path}?info={ex.__str__()}')

            elif action == 'addppl':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_inscripciones')
                    data['title'] = u'Adicionar registro de PPL al Alumno'
                    data['id'] = int(encrypt(request.GET['idi']))
                    data['form'] = PersonaPPLForm()
                    return render(request, "inscripciones/addppl.html", data)
                except Exception as ex:
                    pass

            elif action == 'editppl':
                try:
                    puede_realizar_accion(request, 'sga.puede_modificar_inscripciones')
                    data['title'] = u'Editar registro de PPL al Alumno'
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=int(encrypt(request.GET['idi'])))
                    data['hppl'] = hppl = HistorialPersonaPPL.objects.get(pk=int(encrypt(request.GET['id'])))
                    form = PersonaPPLForm(initial={'fechaingresoppl': hppl.fechaingreso,
                                                   'fechasalidappl': hppl.fechasalida,
                                                   'observacionppl': hppl.observacion,
                                                   'archivoppl': None,
                                                   'centrorehabilitacion': hppl.centrorehabilitacion,
                                                   'lidereducativo': hppl.lidereducativo,
                                                   'correolidereducativo': hppl.correolidereducativo,
                                                   'telefonolidereducativo': hppl.telefonolidereducativo,
                                                   })

                    data['form'] = form
                    return render(request, "inscripciones/editppl.html", data)
                except Exception as ex:
                    pass

            elif action == 'deleteppl':
                try:
                    puede_realizar_accion(request, 'sga.puede_eliminar_ppl')
                    data['title'] = u'Eliminar registro de PPL al Alumno'
                    data['inscripcion'] = Inscripcion.objects.get(pk=request.GET['idi'])
                    data['hppl'] = HistorialPersonaPPL.objects.get(pk=request.GET['id'])
                    return render(request, "inscripciones/deleteppl.html", data)
                except Exception as ex:
                    pass

            elif action == 'listado_ppl':
                try:
                    __author__ = 'Unemi'
                    inscripciones = Inscripcion.objects.filter(status=True, carrera__coordinacion__id__in=[1,2,3,4,5], persona__ppl=True).distinct().order_by( 'persona__apellido1', 'persona__apellido2','persona__nombres')
                    #matriculas = Matricula.objects.filter(status=True,nivel__periodo=periodo, inscripcion__modalidad__id__in = [1,2], inscripcion__carrera__coordinacion__id=9).distinct().order_by( 'inscripcion__persona__apellido1', 'inscripcion__persona__apellido2','inscripcion__persona__nombres')
                    title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('PPL')
                    ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=estudiantes_ppl_' + random.randint(1,10000).__str__() + '.xls'
                    columns = [
                        (u"CEDULA", 3000),
                        (u"APELLIDOS", 5000),
                        (u"NOMBRES", 5000),
                        (u"FACULTAD", 1000),
                        (u"CARRERA", 5000),
                        (u"MODALIDAD", 1000),
                        (u"NIVEL", 500),
                        (u"ESTATUS", 500),
                        (u"CENTRO DE REHABILITACIÓN SOCIAL", 5000),
                        (u"LIDER EDUCATIVO", 5000),
                        (u"CORREO ELECTRONICO", 5000),
                    ]
                    row_num = 4
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    row_num = 5
                    for inscripcion in inscripciones:
                        campo1 = str(inscripcion.persona.identificacion())
                        campo2 = str(inscripcion.persona.apellido1 + " " + inscripcion.persona.apellido2).strip()
                        campo3 = str(inscripcion.persona.nombres)
                        campo4 = str(inscripcion.coordinacion.alias)
                        campo5 = str(inscripcion.carrera.nombre_completo())
                        campo6 = str(inscripcion.modalidad.nombre)
                        campo7 = ""
                        campo8 = "PPL"
                        campo9 = ""
                        campo10 = ""
                        campo11 = ""
                        eInscripcionNivel = inscripcion.mi_nivel()
                        campo7 = str(eInscripcionNivel.nivel.nombre)
                        if inscripcion.persona.historialpersonappl_set.filter().exists():
                            hppl = inscripcion.persona.historialpersonappl_set.filter().order_by('-fechaingreso')[0]
                            campo9 = str(hppl.centrorehabilitacion) if hppl.centrorehabilitacion else ""
                            campo10 = str(hppl.lidereducativo) if hppl.lidereducativo else ""
                            campo11 = str(hppl.correolidereducativo) if hppl.correolidereducativo else ""

                        ws.write(row_num, 0, campo1, font_style2)
                        ws.write(row_num, 1, campo2, font_style2)
                        ws.write(row_num, 2, campo3, font_style2)
                        ws.write(row_num, 3, campo4, font_style2)
                        ws.write(row_num, 4, campo5, font_style2)
                        ws.write(row_num, 5, campo6, font_style2)
                        ws.write(row_num, 6, campo7, font_style2)
                        ws.write(row_num, 7, campo8, font_style2)
                        ws.write(row_num, 8, campo9, font_style2)
                        ws.write(row_num, 9, campo10, font_style2)
                        ws.write(row_num, 10, campo11, font_style2)
                        row_num += 1
                    wb.save(response)
                    return response
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos. %s" % ex})

            elif action == 'listado_matriculado_ppl':
                try:
                    __author__ = 'Unemi'
                    matriculas = Matricula.objects.filter(status=True, nivel__periodo=periodo, inscripcion__carrera__coordinacion__id__in=[1,2,3,4,5], inscripcion__persona__ppl=True).distinct().order_by( 'inscripcion__persona__apellido1', 'inscripcion__persona__apellido2','inscripcion__persona__nombres')
                    title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('PPL')
                    ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=estudiantes_matriculados_ppl_' + random.randint(1,10000).__str__() + '.xls'
                    columns = [
                        (u"CEDULA", 3000),
                        (u"APELLIDOS", 5000),
                        (u"NOMBRES", 5000),
                        (u"FACULTAD", 1000),
                        (u"CARRERA", 5000),
                        (u"MODALIDAD", 1000),
                        (u"MATRICULA", 1000),
                        (u"ESTATUS", 500),
                        (u"CENTRO DE REHABILITACIÓN SOCIAL", 5000),
                        (u"LIDER EDUCATIVO", 5000),
                        (u"CORREO ELECTRONICO", 5000),
                        (u"NIVEL", 500),
                    ]
                    row_num = 4
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    row_num = 5
                    for matricula in matriculas:
                        campo1 = str(matricula.inscripcion.persona.identificacion())
                        campo2 = str(matricula.inscripcion.persona.apellido1 + " " + matricula.inscripcion.persona.apellido2).strip()
                        campo3 = str(matricula.inscripcion.persona.nombres)
                        campo4 = str(matricula.inscripcion.coordinacion.alias)
                        campo5 = str(matricula.inscripcion.carrera.nombre_completo())
                        campo6 = str(matricula.inscripcion.modalidad.nombre)
                        if MateriaAsignada.objects.filter(matricula=matricula, matriculas__gte=2).exists():
                            campo7 = "SEGUNDA"
                        else:
                            campo7 = "PRIMERA"
                        campo8 = "PPL"
                        campo9 = ""
                        campo10 = ""
                        campo11 = ""
                        campo12 = ""
                        if matricula.inscripcion.persona.historialpersonappl_set.filter().exists():
                            hppl = matricula.inscripcion.persona.historialpersonappl_set.filter().order_by('-fechaingreso')[0]
                            campo9 = str(hppl.centrorehabilitacion) if hppl.centrorehabilitacion else ""
                            campo10 = str(hppl.lidereducativo) if hppl.lidereducativo else ""
                            campo11 = str(hppl.correolidereducativo) if hppl.correolidereducativo else ""

                        ws.write(row_num, 0, campo1, font_style2)
                        ws.write(row_num, 1, campo2, font_style2)
                        ws.write(row_num, 2, campo3, font_style2)
                        ws.write(row_num, 3, campo4, font_style2)
                        ws.write(row_num, 4, campo5, font_style2)
                        ws.write(row_num, 5, campo6, font_style2)
                        ws.write(row_num, 6, campo7, font_style2)
                        ws.write(row_num, 7, campo8, font_style2)
                        ws.write(row_num, 8, campo9, font_style2)
                        ws.write(row_num, 9, campo10, font_style2)
                        ws.write(row_num, 10, campo11, font_style2)
                        eInscripcionNivel = matricula.inscripcion.mi_nivel()
                        campo12 = str(eInscripcionNivel.nivel.nombre)
                        ws.write(row_num, 11, campo12, font_style2)
                        row_num += 1
                    wb.save(response)
                    return response
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos. %s" % ex})

            elif action == 'titulacion':
                try:
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])

                    data['title'] = u'Curso complexivo'
                    if 'idperiodogrupo' in request.GET:
                        idperiodogrupo = request.GET['idperiodogrupo']
                        matricula = MatriculaTitulacion.objects.get(Q(inscripcion=inscripcion), (Q(estado=1) | Q(estado=10) | Q(estado=9)), alternativa__grupotitulacion__periodogrupo__id=idperiodogrupo)
                    else:
                        matricula = MatriculaTitulacion.objects.filter(Q(inscripcion=inscripcion), (Q(estado=1) | Q(estado=10) | Q(estado=9)), status=True).order_by('-id')[0]
                    data['matricula'] = matricula
                    data['lstadoperiodos'] = MatriculaTitulacion.objects.values_list('alternativa__grupotitulacion__periodogrupo__id', 'alternativa__grupotitulacion__periodogrupo__nombre').filter(Q(inscripcion=inscripcion), (Q(estado=1) | Q(estado=10) | Q(estado=9))).distinct()
                    data['alternativa'] = matricula.alternativa
                    data['estudiante'] = inscripcion.persona
                    data['materias'] = matricula.alternativa.complexivomateria_set.filter(status=True).order_by('fechainicio')
                    data['archivos'] = ArchivoTitulacion.objects.filter(vigente=True, status=True).distinct()
                    pexamen = 0
                    ppropuesta = 0
                    if matricula.alternativa.cronogramaexamencomplexivo_set.values('id').exists():
                        data['cronograma'] = cronograma = matricula.alternativa.cronogramaexamencomplexivo_set.all()[0]
                        if matricula.alternativa.cronogramaexamencomplexivo_set.get().fechaeleccionpropuestafin != None and matricula.alternativa.cronogramaexamencomplexivo_set.get().fechaeleccionpropuestainicio != None:
                            if datetime.now().date() <= matricula.alternativa.cronogramaexamencomplexivo_set.get().fechaeleccionpropuestafin and datetime.now().date() >= matricula.alternativa.cronogramaexamencomplexivo_set.get().fechaeleccionpropuestainicio:
                                data['disponible'] = True
                            if datetime.now().date() >= matricula.alternativa.cronogramaexamencomplexivo_set.get().fechaeleccionpropuestainicio:
                                data['disponibleinicio'] = True
                    if matricula.alternativa.tiene_examen():

                        if ComplexivoExamenDetalle.objects.filter(status=True, matricula=matricula).exists():
                            listaid = ComplexivoExamenDetalle.objects.values_list('examen_id', flat=True).filter(status=True, matricula=matricula)
                            examenes = ComplexivoExamen.objects.filter(id__in=listaid)
                        else:
                            examenes = matricula.alternativa.complexivoexamen_set.filter(status=True)
                        data['examenes'] = examenes

                        data['detalleexamen'] = detalle = ComplexivoExamenDetalle.objects.filter(status=True, matricula=matricula).order_by('-id')[0]
                        data['examen'] = examen = detalle.examen if detalle else None

                        if detalle:
                            data['detalleexamen'] = detalle
                        else:
                            detalle = ComplexivoExamenDetalle(examen=examen, matricula=matricula)
                            detalle.save(request)
                            log(u"Se creo un detalle de examen complexivo porque no existia %s - [%s] idalter: %s ---creado: %s" % (matricula, examen, matricula.alternativa.id, detalle), request, "delete")
                        pexamen = detalle.ponderacion()
                    listadogrupos = None
                    numgrupos = 0
                    grupo = None
                    idgrupotematica = 0
                    miestadogrupo = None
                    if ComplexivoGrupoTematica.objects.values('id').filter(status=True, complexivodetallegrupo__status=True, complexivodetallegrupo__matricula=matricula, activo=True).exists():
                        listadogrupos = ComplexivoGrupoTematica.objects.filter(status=True, complexivodetallegrupo__status=True, complexivodetallegrupo__matricula=matricula, activo=True).order_by('id')
                        numgrupos = listadogrupos.count()
                        if 'idgrupotematica' in request.GET:
                            idgrupotematica = int(encrypt(request.GET['idgrupotematica']))
                            data['grupo'] = grupo = ComplexivoGrupoTematica.objects.get(pk=idgrupotematica, activo=True)
                        else:
                            data['grupo'] = grupo = ComplexivoGrupoTematica.objects.filter(status=True, complexivodetallegrupo__status=True, complexivodetallegrupo__matricula=matricula, activo=True).order_by('-id')[0]
                            idgrupotematica = grupo.id
                        data['companeros'] = companeros = grupo.complexivodetallegrupo_set.filter(Q(status=True), (Q(matricula__estado=1) | Q(matricula__estado=10) | Q(matricula__estado=9))).exclude(matricula=matricula)
                        if grupo.complexivodetallegrupo_set.filter(Q(status=True), (Q(matricula__estado=1) | Q(matricula__estado=10) | Q(matricula__estado=9)), matricula=matricula):
                            miestadogrupo = grupo.complexivodetallegrupo_set.get(Q(status=True), (Q(matricula__estado=1) | Q(matricula__estado=10) | Q(matricula__estado=9)), matricula=matricula)
                        data['confirmar'] = grupo.complexivodetallegrupo_set.values('id').filter(status=True, estado=4, matricula=matricula).exists()
                        data['tipoarchivo'] = TIPO_ARCHIVO_COMPLEXIVO_PROPUESTA
                        data['propuestas'] = p = grupo.complexivopropuestapractica_set.filter(status=True).order_by('id')
                        data['tutor'] = grupo.tematica.tutor if not grupo.tiene_tematica_confirmar(matricula.inscripcion_id) else ''
                        nintegrantes = companeros.count() + 1
                        data['numerointegrantes'] = grupo.grupocupo.cupoasignado if grupo.grupocupo else nintegrantes
                        data['grupos'] = ComplexivoGrupoTematica.objects.filter(pk__in=ComplexivoDetalleGrupo.objects.values_list("grupo__id", flat=False).filter(matricula=inscripcion.proceso_titulacion(), grupo__complexivoacompanamiento__isnull=False).distinct())
                        ppropuesta = matricula.notapropuestagrupo(grupo)
                    data['idgrupotematica'] = idgrupotematica
                    data['listadogrupos'] = listadogrupos
                    data['numgrupos'] = numgrupos
                    data['miestadogrupo'] = miestadogrupo
                    data['pexamen'] = pexamen
                    data['ppropuesta'] = ppropuesta
                    data['modelos'] = matricula.alternativa.modelo_alternativatitulacion_set.filter(status=True).order_by('modelo__nombre')
                    if matricula.alternativa.tipotitulacion.tipo == 1:
                        data['ptotal'] = ppropuesta
                    if matricula.alternativa.tipotitulacion.tipo == 2:
                        data['ptotal'] = matricula.notafinalcomplexivo(grupo)
                    data['requisitossustentar'] = RequisitoSustentar.objects.filter(status=True)
                    alter = AlternativaTitulacion.objects.get(pk=matricula.alternativa_id)
                    data = valida_matricular_estudiante(data, alter, inscripcion, matricula)
                    return render(request, "inscripciones/viewtitulacion.html", data)
                except Exception as ex:
                    print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
                    return HttpResponseRedirect("/")

            elif action == 'detalle2':
                try:
                    detalles = []
                    if 'idd' in request.GET:
                        detalles = ComplexivoAcompanamiento.objects.get(pk=request.GET['idd'], status=True)
                    data['detalles'] = detalles
                    template = get_template("alu_complexivocurso/detalletutorias.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            elif action == 'detalle':
                try:
                    grupo = []
                    inscripcion = Inscripcion.objects.get(pk=request.GET['idinsc'])

                    if 'ida' in request.GET:
                        grupo = ComplexivoGrupoTematica.objects.get(pk=request.GET['ida'], status=True)
                        tematica = grupo.tematica
                    else:
                        tematica = ComplexivoTematica.objects.get(pk=int(request.GET['id']))
                        if tematica.complexivogrupotematica_set.values('id').filter(status=True).exists():
                            grupo = \
                                ComplexivoGrupoTematica.objects.filter(tematica_id=int(request.GET['id'])).distinct()[0]
                    data['estadoapto'] = int(request.GET['estadoapto'])
                    data['matri'] = MatriculaTitulacion.objects.get(pk=request.GET['idmatri'])
                    data['grupo'] = grupo
                    data['tematica'] = tematica
                    data['idalum'] = inscripcion.id
                    data['grupos'] = ComplexivoGrupoTematica.objects.filter(
                        pk__in=ComplexivoDetalleGrupo.objects.values_list("grupo__id", flat=False).filter(
                            matricula=inscripcion.proceso_titulacion(),
                            grupo__complexivoacompanamiento__isnull=False).distinct())
                    # data['detalles'] = grupo.complexivoacompanamiento_set.filter(status=True).order_by('id')
                    detalles = []
                    if grupo:
                        detalles = grupo.complexivoacompanamiento_set.filter(status=True).order_by('id')
                    data['detalles'] = detalles

                    if 'mostrarcupos' in request.GET:
                        gcupos = tematica.complexivotematicagrupocupo_set.filter(status=True, enuso=False).order_by(
                            'numerogrupo')
                        grupocupos = [[g.numerogrupo, g.cupoasignado] for g in gcupos]
                        data['grupocupos'] = grupocupos

                    template = get_template("alu_complexivocurso/detalle.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            elif action == 'sedeexamenes':
                try:
                    from inno.models import MatriculaSedeExamen
                    data['title'] = u'Sede de exámenes'
                    data['matricula'] = matricula = Matricula.objects.get(pk=request.GET['id'])
                    data['inscripcion'] = matricula.inscripcion
                    data['ret'] = request.GET['ret']
                    data['eMatriculaSedeExamenes'] = MatriculaSedeExamen.objects.filter(status=True, matricula=matricula)
                    form = MatriculaSedeExamenForm()
                    mismaterias = matricula.mismaterias()
                    modeloevaluativos = mismaterias.values_list('materia__modeloevaluativo_id', flat=True) if mismaterias.values("id").exists() else []
                    form.examen(modeloevaluativos)
                    data['form'] = form
                    data['sedes'] = SedeVirtual.objects.filter(sedevirtualperiodoacademico__periodo=periodo, sedevirtualperiodoacademico__status=True, status=True, activa=True).values_list('nombre', 'id').distinct()
                    return render(request, "inscripciones/sedesexamenes/view.html", data)
                except Exception as ex:
                    pass

            elif action == 'addsedeexamen':
                try:
                    from inno.models import MatriculaSedeExamen
                    data['title'] = u"Adicionar sede de examen"
                    data['matricula'] = matricula = Matricula.objects.get(pk=request.GET['id'])
                    data['inscripcion'] = matricula.inscripcion
                    data['ret'] = request.GET['ret']
                    # data['sedes'] = Sede.objects.filter(status=True)
                    form = MatriculaSedeExamenForm()
                    mismaterias = matricula.mismaterias()
                    modeloevaluativos = mismaterias.values_list('materia__modeloevaluativo_id', flat=True) if mismaterias.values("id").exists() else []
                    form.examen(modeloevaluativos)
                    data['form'] = form
                    return render(request, "inscripciones/sedesexamenes/new.html", data)
                except Exception as ex:
                    pass

            elif action == 'editsedeexamen':
                try:
                    from inno.models import MatriculaSedeExamen
                    data['title'] = u"Editar sede de examen"
                    data['eMatriculaSedeExamen'] = eMatriculaSedeExamen = MatriculaSedeExamen.objects.get(pk=request.GET['id'])
                    data['matricula'] = matricula = eMatriculaSedeExamen.matricula
                    data['inscripcion'] = matricula.inscripcion
                    data['ret'] = request.GET['ret']
                    form = MatriculaSedeExamenForm(initial={'sede': eMatriculaSedeExamen.sede,
                                                            'detallemodeloevaluativo': eMatriculaSedeExamen.detallemodeloevaluativo
                                                            }
                                                   )
                    mismaterias = matricula.mismaterias()
                    modeloevaluativos = mismaterias.values_list('materia__modeloevaluativo_id', flat=True) if mismaterias.values("id").exists() else []
                    form.examen(modeloevaluativos)
                    data['form'] = form
                    return render(request, "inscripciones/sedesexamenes/edit.html", data)
                except Exception as ex:
                    pass

            #         Para subir alumnos y asignar de manera masiva

            elif action == 'upload':
                try:
                    data['detalles'] = DetalleModeloEvaluativo.objects.filter(status=True, alternativa_id=20).order_by('id')
                    data['sedes'] = SedeVirtual.objects.filter(status=True).order_by('id')
                    template = get_template("inscripciones/upload_sede_poblacion.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content})
                except Exception as ex:
                    pass

            elif action == 'loaddatesede':
                try:
                    if not 'sede' in request.GET:
                        raise NameError(u"Debe seleccionar una sede")
                    sede = request.GET['sede']
                    fechas = list(FechaPlanificacionSedeVirtualExamen.objects.filter(sede_id=sede, periodo=periodo).values_list('id', 'fecha'))
                    return JsonResponse({"result": "ok", 'fechas': fechas})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos. %s" % ex.__str__()})
            elif action == 'loadturnsede':
                try:
                    if not 'fecha' in request.GET:
                        raise NameError(u"Debe seleccionar una fecha")
                    fecha = request.GET['fecha']
                    eFecha = FechaPlanificacionSedeVirtualExamen.objects.filter(status=True, id=fecha).first()
                    if not eFecha:
                        raise NameError(u"Debe seleccionar una fecha")
                    sede = eFecha.sede_id
                    turnosexclude = json.loads(request.GET['turnosselect']) if not sede == 11 else []
                    turnos = list(TurnoPlanificacionSedeVirtualExamen.objects.filter(fechaplanificacion_id=fecha).exclude(id__in=turnosexclude).filter(id__in=turnos_disponibles_fecha(fecha)).values_list('id', 'horainicio', 'horafin'))
                    return JsonResponse({"result": "ok", 'turnos': turnos})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos. %s" % ex.__str__()})
            elif action == 'loadcoursesede':
                try:
                    if not 'turno' in request.GET:
                        raise NameError(u"Debe seleccionar un turno")
                    turno = request.GET['turno']
                    aulas = []
                    for aula in AulaPlanificacionSedeVirtualExamen.objects.filter(turnoplanificacion_id=turno):
                        cupo_disponible=aula.cupo_disponible()
                        aulas.append([aula.id, '{} ({})'.format(aula.aula.nombre, cupo_disponible), int(cupo_disponible)])
                    # aulas = list(AulaPlanificacionSedeVirtualExamen.objects.filter(turnoplanificacion_id=turno).values_list('id', 'aula__nombre'))
                    return JsonResponse({"result": "ok", 'aulas': aulas})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos. %s" % ex.__str__()})

            elif action == 'titulacioncarrera':
                try:
                    periodo = request.session['periodo']
                    data['configpormalla'] = False
                    data['carreralist'] = Carrera.objects.filter(status=True, coordinacion__lte=5,coordinadorcarrera__periodo=periodo)
                    template = get_template("inscripciones/modal_titulacioncarrera.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": "Error al obtener los datos."})

            elif action == 'generareportetitulacion':
                try:
                    periodo = request.session['periodo']
                    carrera = request.GET.get('idcar')
                    tiporeporte = int(request.GET.get('tiporeporte'))
                    tipomodalidad = int(request.GET.get('tipomodalidad'))
                    modalidad = 'EXAMEN COMPLEXIVO' if tipomodalidad == 1 else 'PROYECTOS INTEGRADORES'
                    reporte = 'Excel' if tiporeporte == 1 else 'Pdf'
                    idmalla = InscripcionMalla.objects.filter(status=True, malla__carrera=carrera,malla__vigente=True).values_list('malla_id',flat=True).distinct('malla_id')
                    malla = Malla.objects.get(pk=idmalla[0])
                    materias = Materia.objects.filter(status=True,asignaturamalla__validarequisitograduacion=True, nivel__periodo=periodo,  asignaturamalla__malla__carrera=carrera,asignaturamalla__malla=malla,asignaturamalla__nivelmalla=malla.niveles_regulares).values_list('id', flat=True)
                    titulacion = MateriaTitulacion.objects.filter(status=True,actacerrada=False,numeroacta=0,estadograduado=False, materiaasignada__materia__in=materias,materiaasignada__matricula__inscripcion__carrera=carrera)
                    if not titulacion:
                        """Validacion Malla: UNIDAD TITULACION"""
                        materias = Materia.objects.filter(status=True,asignaturamalla__malla=383,asignatura=10457,asignaturamalla__validarequisitograduacion=True,
                                                          nivel__periodo=periodo,asignaturamalla__malla__carrera=208,asignaturamalla__nivelmalla=1,
                                                          materiaasignada__matricula__inscripcion__carrera=carrera).values_list('id',flat=True).distinct()
                        titulacion = MateriaTitulacion.objects.filter(status=True,actacerrada=False,numeroacta=0,estadograduado=False, materiaasignada__materia__in=materias,materiaasignada__matricula__inscripcion__carrera=carrera)
                        if not titulacion:
                            return JsonResponse({"result": False, "mensaje": f'No existe inscritos para titulacion'})

                    notifi = Notificacion(cuerpo=f'Generación de reporte en {reporte}',
                                          titulo='Reporte de apto/no apto para titulacion',
                                          destinatario=personasesion,
                                          url='',
                                          prioridad=1, app_label='SGA',
                                          fecha_hora_visible=datetime.now() + timedelta(days=1), tipo=2,
                                          en_proceso=True)
                    notifi.save(request)
                    reporte_titulacionporcarrera(request=request, notiid=notifi.id, periodo=periodo,materias=materias,carrera=carrera,tipomodalidad=modalidad, malla=malla, tiporeporte=tiporeporte).start()
                    return JsonResponse({"result": True,
                                         "mensaje": u"El reporte de estudiantes aptos/no aptos para titulacion se está realizando. Verifique su apartado de notificaciones después de unos minutos.",
                                         "btn_notificaciones": traerNotificaciones(request, data, persona)})
                except Exception as ex:
                    return JsonResponse({"result": False,
                                         "mensaje":"Hubo un error al generar el reporte"})
            return HttpResponseRedirect(request.path)
        else:
            try:
                url_vars = ''
                facultad_admision = 9
                data['title'] = u'Listado de inscripciones'
                persona = request.session['persona']
                carreras = Carrera.objects.filter(status=True).exclude(coordinacion__id=9)
                search = None
                if 'mensj' in request.GET:
                    data['mensj'] = request.GET['mensj']
                if 'info' in request.GET:
                    data['info'] = request.GET['info']
                ids = None
                menserror = 0
                if 'menserror' in request.GET:
                    menserror = request.GET['menserror']
                data['menserror'] = menserror
                bandera = 0
                filtros = Q(status=True)
                # inscripciones = Inscripcion.objects.select_related().filter(status=True).exclude(carrera__coordinacion__id=9)
                if 'id' in request.GET:
                    bandera = 1
                    ids = request.GET['id']
                    url_vars += f'&id={ids}'
                    # inscripciones = inscripciones.filter(id=ids)
                    filtros = filtros & Q(id=ids)
                if 's' in request.GET:
                    bandera = 1
                    search = request.GET['s'].strip()
                    url_vars += f'&s={search}'
                    ss = search.split(' ')
                    if len(ss) == 1:
                        # inscripciones = inscripciones.filter(Q(persona__nombres__icontains=search) |
                        #                                      Q(persona__apellido1__icontains=search) |
                        #                                      Q(persona__apellido2__icontains=search) |
                        #                                      Q(persona__cedula__icontains=search) |
                        #                                      Q(persona__pasaporte__icontains=search) |
                        #                                      Q(identificador__icontains=search) |
                        #                                      Q(inscripciongrupo__grupo__nombre__icontains=search) |
                        #                                      Q(persona__usuario__username__icontains=search))
                        filtros = filtros & Q(Q(persona__nombres__icontains=search) |
                                                             Q(persona__apellido1__icontains=search) |
                                                             Q(persona__apellido2__icontains=search) |
                                                             Q(persona__cedula__icontains=search) |
                                                             Q(persona__pasaporte__icontains=search) |
                                                             Q(identificador__icontains=search) |
                                                             Q(inscripciongrupo__grupo__nombre__icontains=search) |
                                                             Q(persona__usuario__username__icontains=search))
                    else:
                        # inscripciones = inscripciones.filter(Q(persona__apellido1__icontains=ss[0]) &
                        #                                      Q(persona__apellido2__icontains=ss[1]))
                        filtros = filtros & Q(Q(persona__apellido1__icontains=ss[0]) & Q(persona__apellido2__icontains=ss[1]))
                carreraselect = 0
                if 'c' in request.GET:
                    carreraselect = int(request.GET['c'])
                    url_vars += f'&c={carreraselect}'
                    if carreraselect > 0:
                        bandera = 1
                        # inscripciones = inscripciones.filter(carrera_id=carreraselect)
                        filtros = filtros & Q(carrera_id=carreraselect)
                modalidadselect = 0
                if 'm' in request.GET:
                    modalidadselect = int(request.GET['m'])
                    url_vars += f'&m={modalidadselect}'
                    if modalidadselect > 0:
                        bandera = 1
                        # inscripciones = inscripciones.filter(modalidad_id=modalidadselect)
                        filtros = filtros & Q(modalidad_id=modalidadselect)
                ppl = 0
                if 'ppl' in request.GET and not request.GET['ppl'] == 'undefined' and int(request.GET['ppl']) > 0:
                    bandera = 1
                    ppl = int(request.GET['ppl'])
                    url_vars += f'&ppl={ppl}'
                    # inscripciones = inscripciones.filter(persona__ppl=int(request.GET['ppl']) == 1)
                    filtros = filtros & Q(persona__ppl=int(request.GET['ppl']) == 1)
                inscripciones = Inscripcion.objects.prefetch_related().filter(filtros).exclude(carrera__coordinacion__id=9).only('id')
                paging = MiPaginador(inscripciones, 8)
                p = 1
                try:
                    paginasesion = 1
                    if 'paginador' in request.session:
                        paginasesion = int(request.session['paginador'])
                    if 'page' in request.GET:
                        p = int(request.GET['page'])
                    else:
                        p = paginasesion
                    try:
                        page = paging.page(p)
                    except:
                        p = 1
                    page = paging.page(p)
                except:
                    page = paging.page(p)
                request.session['paginador'] = p
                data['paging'] = paging
                ids_inscripciones_pagina = [inscripcion.id for inscripcion in page.object_list]
                inscripciones_con_anotacion = Inscripcion.objects.filter(id__in=ids_inscripciones_pagina).\
                    annotate(es_graduado=Exists(Graduado.objects.filter(status=True, estadograduado=True, inscripcion_id=OuterRef('id'))),
                             es_egresado=Exists(Egresado.objects.filter(status=True, inscripcion_id=OuterRef('id'))),
                             es_retirocarrera=Exists(RetiroCarrera.objects.filter(status=True, inscripcion_id=OuterRef('id'))),
                             activo_perfilinscripcion=Exists(PerfilUsuario.objects.filter(status=True, visible=True, inscripcion_id=OuterRef('id'))))
                data['rangospaging'] = paging.rangos_paginado(p)
                data['page'] = page
                data['search'] = search if search else ""
                data['ids'] = ids if ids else ""
                data['inscripciones'] = inscripciones_con_anotacion
                data['utiliza_grupos_alumnos'] = UTILIZA_GRUPOS_ALUMNOS
                data['clave'] = DEFAULT_PASSWORD
                data['reporte_0'] = obtener_reporte('listado_vinculacion_estudiante')
                data['reporte_1'] = obtener_reporte('ficha_inscripcion')
                data['reporte_2'] = obtener_reporte('certificado_inscripcion')
                data['reporte_3'] = obtener_reporte('certificado_matricula_alumno_plus')
                data['reporte_4'] = obtener_reporte('hoja_vida_sagest')
                data['reporte_5'] = obtener_reporte('certificado_matricula_alumno')
                data['reporte_6'] = obtener_reporte("record_alumno")
                data['reporte_7'] = obtener_reporte("registro_matriculacion")
                data['reporte_8'] = obtener_reporte("compromiso_pago")
                data['control_unico_credenciales'] = CONTROL_UNICO_CREDENCIALES
                data['usa_tipos_inscripciones'] = USA_TIPOS_INSCRIPCIONES
                data['matriculacion_libre'] = variable_valor('MATRICULACION_LIBRE')
                data['add_inscripcion'] = variable_valor('ADICIONAR_INSCRIPCION')
                data['inscribir_otracarrera'] = variable_valor('INSCRIBIR_OTRA_CARRERA')
                data['periodo'] = request.session['periodo']
                data['carreraselect'] = carreraselect
                data['modalidadselect'] = modalidadselect
                data['carreras'] = carreras
                data['pers'] = persona
                data['ppl'] = ppl
                data['url_vars'] = url_vars
                return render(request, "inscripciones/view.html", data)
            except Exception as ex:
                pass

def valida_matricular_estudiante(data, alter, inscripcion, matriculatitulacion):
    vali_alter = 8
    vali_tenido = 0
    data['item'] = alter
    data['grupotitulacion'] = alter.grupotitulacion
    data['inscripcion'] = inscripcion
    malla = inscripcion.inscripcionmalla_set.filter(status=True)[0].malla
    perfil = inscripcion.persona.mi_perfil()
    fechainicioprimernivel = inscripcion.fechainicioprimernivel if inscripcion.fechainicioprimernivel else datetime.now().date()
    excluiralumnos = datetime(2009, 1, 21, 23, 59, 59).date()
    data['esexonerado'] = fechainicioprimernivel <= excluiralumnos

    data['tiene_discapidad'] = perfil.tienediscapacidad
    # if alter.estadofichaestudiantil:
    #     vali_alter += 1
    ficha = 0
    if inscripcion.persona.nombres and inscripcion.persona.apellido1 and inscripcion.persona.nacimiento and  inscripcion.persona.nacionalidad and inscripcion.persona.email and inscripcion.persona.estado_civil and inscripcion.persona.sexo:
        data['datospersonales'] = True
        ficha += 1
    if inscripcion.persona.paisnacimiento and inscripcion.persona.provincianacimiento and inscripcion.persona.cantonnacimiento and inscripcion.persona.parroquianacimiento:
        data['datosnacimientos'] = True
        ficha += 1
    examenfisico = inscripcion.persona.datos_examen_fisico()
    if inscripcion.persona.sangre and examenfisico.peso and examenfisico.talla:
        data['datosmedicos'] = True
        ficha += 1
    if inscripcion.persona.pais and inscripcion.persona.provincia and inscripcion.persona.canton and inscripcion.persona.parroquia and inscripcion.persona.direccion and inscripcion.persona.direccion2 and inscripcion.persona.num_direccion and inscripcion.persona.telefono_conv or inscripcion.persona.telefono:
        data['datosdomicilio'] = True
        ficha += 1
    if perfil.raza:
        data['etnia'] = True
        ficha += 1
    if ficha == 5:
        vali_tenido += 1
    # if alter.estadopracticaspreprofesionales:
    #     vali_alter += 1
    totalhoras = 0
    practicaspreprofesionalesinscripcion = PracticasPreprofesionalesInscripcion.objects.filter(inscripcion=inscripcion, status=True, culminada=True)
    data['malla_horas_practicas'] = malla.horas_practicas
    if fechainicioprimernivel > excluiralumnos:
        if practicaspreprofesionalesinscripcion.exists():
            for practicas in practicaspreprofesionalesinscripcion:
                if practicas.tiposolicitud == 3:
                    totalhoras += practicas.horahomologacion if practicas.horahomologacion else 0
                else:
                    totalhoras += practicas.numerohora
            if totalhoras >= malla.horas_practicas:
                data['practicaspreprofesionales'] = True
                vali_tenido += 1
        data['practicaspreprofesionalesvalor'] = totalhoras
    else:
        data['practicaspreprofesionales'] = True
        vali_tenido += 1
        data['practicaspreprofesionalesvalor'] = malla.horas_practicas
    # if alter.estadocredito:
    #     vali_alter += 1
    data['creditos'] = inscripcion.aprobo_asta_penultimo_malla()
    if inscripcion.aprobo_asta_penultimo_malla() and inscripcion.esta_matriculado_ultimo_nivel():
        vali_tenido += 1
    data['cantasigaprobadas'] = inscripcion.cantidad_asig_aprobada_penultimo_malla()
    data['cantasigaprobar'] = inscripcion.cantidad_asig_aprobar_penultimo_malla()
    data['esta_mat_ultimo_nivel'] = inscripcion.esta_matriculado_ultimo_nivel()
    # if alter.estadoadeudar:
    #     vali_alter += 1
    if inscripcion.adeuda_a_la_fecha() == 0:
        data['deudas'] = True
        vali_tenido += 1
    data['deudasvalor'] = inscripcion.adeuda_a_la_fecha()
    # if alter.estadoingles:
    #     vali_alter += 1
    modulo_ingles = ModuloMalla.objects.filter(malla=malla, status=True).exclude(asignatura_id=782)
    numero_modulo_ingles = modulo_ingles.count()
    lista = []
    listaid = []
    for modulo in modulo_ingles:
        if inscripcion.estado_asignatura(modulo.asignatura) == NOTA_ESTADO_APROBADO:
            lista.append(modulo.asignatura.nombre)
            listaid.append(modulo.asignatura.id)
    data['modulo_ingles_aprobados'] = lista
    data['modulo_ingles_faltante'] = modulo_ingles.exclude(asignatura_id__in=[int(i) for i in listaid])
    if numero_modulo_ingles == len(listaid):
        data['modulo_ingles'] = True
        vali_tenido += 1
    # if alter.estadonivel:
    #     vali_alter += 1
    #total_materias_malla = malla.cantidad_materiasaprobadas()
    #cantidad_materias_aprobadas_record = inscripcion.recordacademico_set.filter(aprobada=True, status=True,asignatura__in=[x.asignatura for x in malla.asignaturamalla_set.filter(status=True)]).count()
    #poraprobacion = round(cantidad_materias_aprobadas_record * 100 / total_materias_malla, 0)
    data['mi_nivel'] = nivel = inscripcion.mi_nivel()
    inscripcionmalla = inscripcion.malla_inscripcion()
    niveles_maximos = inscripcionmalla.malla.niveles_regulares
    # MOMENTANEO EL AUMENTO DE VALI_TENIDO
    # vali_tenido += 1

    asignaturas_malla = AsignaturaMalla.objects.filter(Q(status=True,malla=malla,opcional=False)| Q(itinerario__in=[1,2,3],status=True,malla=malla,opcional=False))
    materia_aprobada = True
    for asignatura in asignaturas_malla:
        existe_asignatura =RecordAcademico.objects.filter(status=True,asignaturamalla=asignatura, inscripcion= inscripcion).exists()
        if existe_asignatura:
            asignatura_record =RecordAcademico.objects.get(status=True,asignaturamalla=asignatura, inscripcion= inscripcion)
            if not asignatura_record.aprobada:
                materia_aprobada = False
                break
        else:
            materia_aprobada = False
            break


    # if poraprobacion >= 100 :
    #     data['nivel'] = True
    #     vali_tenido += 1

    if materia_aprobada:
        data['nivel'] = True
        vali_tenido += 1

    else:
        if niveles_maximos == nivel.nivel.id:
            data['septimo'] = True
    if perfil.tienediscapacidad:
        data['discapacidad'] = perfil
    if inscripcion.persona.sexo.id == ESTADO_GESTACION:
        data['femenino'] = True
    # if alter.estadovinculacion:
    #     vali_alter += 1
    data['malla_horas_vinculacion'] = malla.horas_vinculacion
    # horastotal = ParticipantesMatrices.objects.filter(matrizevidencia_id=2, status=True, proyecto__status=True, inscripcion_id=inscripcion.id, actividad__isnull=True).aggregate(horastotal=Sum('horas'))['horastotal']
    horastotal = null_to_numeric(ParticipantesMatrices.objects.filter(matrizevidencia_id=2, status=True, proyecto__status=True, inscripcion_id=inscripcion.id, actividad__isnull=True).aggregate(horastotal=Sum('horas'))['horastotal'])
    horasconvalidadas = null_to_numeric(inscripcion.participantesmatrices_set.filter(matrizevidencia_id=2, status=True, actividad__isnull=False).aggregate(horas=Sum('horas'))['horas'])

    # horastotal = horastotal if horastotal else 0
    horastotal += horasconvalidadas
    if fechainicioprimernivel > excluiralumnos:
        if horastotal >= malla.horas_vinculacion:
            data['vinculacion'] = True
            vali_tenido += 1
        data['horas_vinculacion'] = horastotal
    else:
        data['horas_vinculacion'] = malla.horas_vinculacion
        data['vinculacion'] = True
        vali_tenido += 1
    # if alter.estadocomputacion:
    #     vali_alter += 1
    asignatura = AsignaturaMalla.objects.values_list('asignatura__id', flat=True).filter(malla__id=32)
    data['record_computacion'] = record = RecordAcademico.objects.filter(inscripcion__id=inscripcion.id, asignatura__id__in=asignatura, aprobada=True)
    creditos_computacion = 0
    data['malla_creditos_computacion'] = malla.creditos_computacion
    for comp in record:
        creditos_computacion += comp.creditos
    if creditos_computacion >= malla.creditos_computacion:
        data['computacion'] = True
        vali_tenido += 1
    data['creditos_computacion'] = creditos_computacion
    if vali_alter == vali_tenido:
        data['aprueba'] = True
        matriculatitulacion.cumplerequisitos = 2
        matriculatitulacion.save()
    else:
        matriculatitulacion.cumplerequisitos = 3
        matriculatitulacion.save()
    if inscripcion.persona.tipocelular == 0:
        data['tipocelular'] = '-'
    else:
        data['tipocelular'] = TIPO_CELULAR[int(inscripcion.persona.tipocelular) - 1][1]
    return data

def turnos_disponibles_fecha(fecha_id):
    turnos_= []
    for turno in TurnoPlanificacionSedeVirtualExamen.objects.filter(fechaplanificacion_id=fecha_id):
        aulas_ = 0
        for aula in turno.aulaplanificacionsedevirtualexamen_set.filter(status=True):
            # Si encuentra al menos un aula con cupo
            if aula.cupo_disponible() > 0:
                aulas_+=1
        if aulas_>0:
            turnos_.append(turno.id)
    return turnos_


