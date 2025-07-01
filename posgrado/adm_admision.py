# -*- coding: UTF-8 -*-
import json
import os
from itertools import count
import random
import time
import pyqrcode
import xlrd
from decimal import Decimal
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Sum, Case, When
from operator import itemgetter
from django.http import HttpResponseRedirect, JsonResponse, HttpResponse
from django.shortcuts import render
from django.template.loader import get_template
from django.template.context import Context
from django.db.models.query_utils import Q
from datetime import datetime, timedelta
from django.db import connections
from django.db.models.aggregates import Count, Max
from django.contrib import messages

from xlwt import *
from xlwt import easyxf
import xlwt

from django.forms import model_to_dict
from decorators import secure_module, last_access
from sagest.models import Rubro, TipoOtroRubro, ComprobanteAlumno, DistributivoPersona
from settings import USA_TIPOS_INSCRIPCIONES, TIPO_INSCRIPCION_INICIAL, SITE_STORAGE
from sga.commonviews import adduserdata, obtener_reporte, actualizar_nota_grupo, copiar_nota_entrevista
from posgrado.forms import AdmiPeriodoForm, CohorteMaestriaForm, AdmiRequisitosMaestriaForm, AdmiPreguntasMaestriaForm, \
    RequisitoForm, EntrevistadorCohorteForm, GrupoExamenForm, PreguntaForm, \
    TipoPreguntaForm, GrupoEntrevistaForm, TablaEntrevistaMaestriaForm, AdmiPagoExamenForm, AdmiPagoMatriculaForm, \
    MatriculaPagoCuota, MatrizInscritosEntrevistaForm, MatrizInscritosEntrevistaNotasForm, InscripcionCarreraForm, \
    FormatoCarreraForm, PreInscripcionForm, RegistroRequisitosMaestriaForm, EvidenciasMaestriasForm, \
    RequisitosMaestriaForm, RequisitosMaestriaImgForm, ConfigurarFirmaAdmisionPosgradoForm, ClasificacionRequisitoForm, \
    TipoClasificacionReqForm, ContratoPagoMaestriaForm, TipoPersonaForm, CopiarRequisitosForm
from posgrado.models import Pago, RequisitosMaestria, PreguntaMaestria, InscripcionCohorte, CohorteMaestria, \
    MaestriasAdmision, DetalleAprobacionContrato, \
    Requisito, GrupoEntrevistaMsc, IntegranteGrupoExamenMsc, IntegranteGrupoEntrevitaMsc, Requisito, GrupoExamenMsc, \
    PreguntasPrograma, TipoPreguntasPrograma, RespuestaEntrevitaMsc, EstadoEntrevista, EvidenciaRequisitosAspirante, \
    DetalleEvidenciaRequisitosAspirante, \
    TablaEntrevistaMaestria, CuotaPago, CuotaPagadas, InscripcionAspirante, EvidenciaPagoExamen, PreInscripcion, \
    FormatoCarreraIpec, EvidenciasMaestrias, RequisitosGrupoCohorte, ConfigurarFirmaAdmisionPosgrado, ClaseRequisito, \
    TipoClasificacionRequisito, Contrato, TablaAmortizacion, TipoPersonaRequisito, DetallePreAprobacionPostulante
from sga.funciones import MiPaginador, log, generar_nombre, convertir_fecha, convertir_hora, variable_valor, \
    resetear_clave, puede_realizar_accion, puede_ver_todoadmision, resetear_clavepostulante, null_to_decimal, \
    validar_archivo, notificacion3
from sga.models import Modalidad, ESTADO_REVISION, miinstitucion, AsignaturaMalla, Persona, Inscripcion, \
    DocumentosDeInscripcion, InscripcionTesDrive, InscripcionTipoInscripcion, CUENTAS_CORREOS, Carrera, Profesor, \
    Coordinacion, CoordinacionImagenes, FirmaPersona, Malla, Administrativo, ItinerarioMallaEspecilidad, Reporte
from inno.models import ProgramaPac, InfraestructuraEquipamientoInformacionPac, TipoFormaPagoPac
from sga.funciones_templatepdf import contratoformapagoprograma
from sga.templatetags.sga_extras import encrypt
from sga.funcionesxhtml2pdf import conviert_html_to_pdf, conviert_html_to_pdfsaveqrcertificado
from sga.tasks import send_html_mail, conectar_cuenta
from moodle import moodle
import shutil
from sga.reportes import run_report_v1
from posgrado.commonviews import secuencia_contratopagare

@login_required(redirect_field_name='ret', login_url='/loginposgrado')
@secure_module
@last_access
@transaction.atomic()
def view(request):
    data = {}
    adduserdata(request, data)
    data['personasesion'] = persona = request.session['persona']
    perfilprincipal = request.session['perfilprincipal']
    hoy = datetime.now().date()

    periodo = request.session['periodo']
    if request.method == 'POST':
        action = request.POST['action']

        if action == 'addmaestriaadmision':
            try:
                f = AdmiPeriodoForm(request.POST)
                if f.is_valid():
                    if ProgramaPac.objects.filter(carrera=f.cleaned_data['carrera'], status=True).exists():
                        programapac = ProgramaPac.objects.filter(carrera=f.cleaned_data['carrera'], status=True).last()
                        if programapac.funcionsustantivadocenciapac_set.filter(status=True).exists():
                            funcion = programapac.funcionsustantivadocenciapac_set.filter(status=True).last()
                            if not funcion.detalleperfilingreso_set.filter(status=True).exists():
                                return JsonResponse({"result": "bad",
                                                     "mensaje": u"Para poder adicionar el programa, primero debe registrar el Perfil de ingreso de la carrera en el módulo de Instauración(Diseño o ajuste de la carrera)."})
                        else:
                            return JsonResponse({"result": "bad",
                                                 "mensaje": u"Para poder adicionar el programa, primero debe registrar el Perfil de ingreso de la carrera en el módulo de Instauración(Diseño o ajuste de la carrera)."})
                    else:
                        return JsonResponse({"result": "bad",
                                             "mensaje": u"Para poder adicionar el programa, primero debe registrar la carrera con su perfil de ingreso en el módulo de Instauración(Diseño o ajuste de la carrera)."})

                    if not MaestriasAdmision.objects.filter(carrera=f.cleaned_data['carrera'], status=True).exists():
                        maestriaadmision = MaestriasAdmision(carrera=f.cleaned_data['carrera'],
                                                             descripcion=f.cleaned_data['descripcion'])
                        maestriaadmision.save(request)
                        log(u'Adicionó Maestría Admisión: %s' % maestriaadmision, request, "add")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        return JsonResponse({"result": "bad", "mensaje": u"La Maestría de Admisión esta ingresado."})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'addpagoexamen':
            try:
                f = AdmiPagoExamenForm(request.POST)
                if f.is_valid():
                    if not Pago.objects.filter(inscripcioncohorte_id=int(encrypt(request.POST['id'])), tipo_id=2,
                                               status=True):
                        pagos = Pago(inscripcioncohorte_id=int(encrypt(request.POST['id'])), tipo_id=2)
                        pagos.save(request)
                        detallepagos = CuotaPago(pago=pagos, fechapago=f.cleaned_data['fecha'],
                                                 valor=f.cleaned_data['valor'])
                        detallepagos.save(request)
                        log(u'Adicionó pago: %s' % pagos, request, "add")
                        return JsonResponse({"result": "ok"})
                    else:
                        if f.cleaned_data['valor'] == 0:
                            pago = Pago.objects.get(inscripcioncohorte_id=int(encrypt(request.POST['id'])), tipo_id=2,
                                                    status=True)
                            log(u'Eliminó pago: %s' % pago, request, "del")
                            pago.status = False
                            pago.save()
                        else:
                            cuota = CuotaPago.objects.get(pago__inscripcioncohorte_id=int(encrypt(request.POST['id'])),
                                                          pago__tipo_id=2, status=True)
                            cuota.fechapago = f.cleaned_data['fecha']
                            cuota.valor = f.cleaned_data['valor']
                            cuota.save(request)
                            log(u'Editó pago: %s' % cuota.pago.inscripcioncohorte, request, "edit")
                        return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'addpagocuota':
            try:
                f = MatriculaPagoCuota(request.POST)
                if f.is_valid():
                    tablepagos = Pago.objects.get(pk=int(encrypt(request.POST['id'])))
                    listacuotas = Pago.objects.filter(inscripcioncohorte=tablepagos.inscripcioncohorte, tipo_id=3,
                                                      cancelado=False, status=True).order_by('numerocuota')
                    cuotapago = CuotaPago(pago=tablepagos,
                                          valor=f.cleaned_data['valor'],
                                          fechapago=f.cleaned_data['fecha'])
                    cuotapago.save(request)
                    valorpagar = f.cleaned_data['valor']
                    cuotafija = tablepagos.inscripcioncohorte.cohortes.valorcuota
                    for lista in listacuotas:
                        totalcancelado = 0
                        if CuotaPagadas.objects.filter(pago=lista, cancelado=False, status=True):
                            totalcancelado = \
                                CuotaPagadas.objects.filter(pago=lista, cancelado=False, status=True).aggregate(
                                    valor=Sum('valor'))['valor']
                        if valorpagar > 0:
                            if totalcancelado > 0:
                                faltante = cuotafija - totalcancelado
                                if valorpagar >= faltante:
                                    CuotaPagadas.objects.filter(pago=lista, cancelado=False, status=True).update(
                                        cancelado=True)
                                    cuotaspagadas = CuotaPagadas(
                                        pago=lista,
                                        cuotapago=cuotapago,
                                        cancelado=True,
                                        valor=faltante,
                                        status=True)
                                    cuotaspagadas.save(request)
                                    lista.cancelado = True
                                    lista.save(request)
                                    if valorpagar == faltante:
                                        break
                                    valorpagar = valorpagar - faltante
                                else:
                                    cuotaspagadas = CuotaPagadas(
                                        pago=lista,
                                        cancelado=False,
                                        cuotapago=cuotapago,
                                        valor=valorpagar,
                                        status=True)
                                    cuotaspagadas.save(request)
                                    break
                            else:
                                if valorpagar >= cuotafija:
                                    cuotaspagadas = CuotaPagadas(
                                        pago=lista,
                                        cuotapago=cuotapago,
                                        cancelado=True,
                                        valor=cuotafija,
                                        status=True)
                                    cuotaspagadas.save(request)
                                    lista.cancelado = True
                                    lista.save(request)
                                else:
                                    if valorpagar == 0:
                                        break
                                    else:
                                        cuotaspagadas = CuotaPagadas(
                                            pago=lista,
                                            cuotapago=cuotapago,
                                            cancelado=False,
                                            valor=valorpagar,
                                            status=True)
                                        cuotaspagadas.save(request)
                                valorpagar = valorpagar - cuotafija

                    # log(u'Adicionó pago: %s' % pagos, request,"add")
                    return JsonResponse({"result": "ok"})

                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'matrizlistaexamen':
            try:
                import openpyxl
                a = 0
                if 'archivomatriz' in request.FILES:
                    newfile = request.FILES['archivomatriz']
                    if newfile.size > 2194304:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Error, Tamaño de archivo Maximo permitido es de 2Mb"})
                cohorte = CohorteMaestria.objects.get(status=True, pk=int(encrypt(request.POST['id'])))
                f = MatrizInscritosEntrevistaForm(request.POST)
                if f.is_valid():
                    newfile = None
                    if 'archivomatriz' in request.FILES:
                        newfile = request.FILES['archivomatriz']
                        newfile._name = generar_nombre("archivomatriz_", newfile._name)
                        cohorte.archivomatriz = newfile
                        cohorte.save(request)
                    # miarchivo = openpyxl.load_workbook(request.FILES['archivomatriz'])
                    miarchivo = xlrd.open_workbook(cohorte.archivomatriz.file.name)
                    # hojas = miarchivo.get_sheet_names()
                    hojas = miarchivo.sheet_by_index(0)
                    # lista = miarchivo.get_sheet_by_name(str(hojas[0]))
                    lista = range(hojas.nrows)
                    totallista = hojas.nrows
                    notapasarexamen = int(cohorte.notaminimaexa)
                    for listadoarchivo in range(hojas.nrows):
                        a += 1
                        if a > 1:
                            filas = hojas.row_values(listadoarchivo)
                            if IntegranteGrupoExamenMsc.objects.filter(grupoexamen__cohorte=cohorte,
                                                                       inscripcion__inscripcionaspirante__persona__cedula=str(
                                                                           filas[0])).exists():
                                integrante = IntegranteGrupoExamenMsc.objects.get(grupoexamen__cohorte=cohorte,
                                                                                  inscripcion__inscripcionaspirante__persona__cedula=str(
                                                                                      filas[0]))
                                integrante.notaexa = filas[4]
                                integrante.notatest = filas[5]
                                integrante.notafinal = filas[6]
                                if filas[6] >= notapasarexamen:
                                    integrante.estado = 2
                                if filas[6] < notapasarexamen:
                                    integrante.estado = 3
                                integrante.save(request)
                    return JsonResponse({'result': 'ok'})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': u'Error al guardar los datos'})

        elif action == 'matrizlistanotaentrevista':
            try:
                a = 0
                if 'archivomatriz' in request.FILES:
                    newfile = request.FILES['archivomatriz']
                    if newfile.size > 2194304:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Error, Tamaño de archivo Maximo permitido es de 2Mb"})
                cohorte = CohorteMaestria.objects.get(status=True, pk=int(encrypt(request.POST['id'])))
                notaaprobar = int(cohorte.notaminimatest)
                f = MatrizInscritosEntrevistaNotasForm(request.POST)
                if f.is_valid():
                    miarchivo = xlrd.open_workbook(file_contents=request.FILES['archivomatriz'].read())
                    hojas = miarchivo.sheet_by_index(0)
                    # lista = miarchivo.get_sheet_by_name(str(hojas[0]))
                    lista = range(hojas.nrows)
                    totallista = hojas.nrows
                    for listadoarchivo in range(hojas.nrows):
                        a += 1
                        if a > 1:
                            filas = hojas.row_values(listadoarchivo)
                            if IntegranteGrupoEntrevitaMsc.objects.filter(grupoentrevista__cohortes=cohorte,
                                                                          inscripcion__inscripcionaspirante__persona__cedula=str(
                                                                              filas[0])).exists():
                                integranteentrevista = IntegranteGrupoEntrevitaMsc.objects.get(
                                    grupoentrevista__cohortes=cohorte,
                                    inscripcion__inscripcionaspirante__persona__cedula=str(filas[0]))
                                notalfinalexamen = \
                                    integranteentrevista.inscripcion.integrantegrupoexamenmsc_set.filter(status=True)[
                                        0].notafinal
                                integranteentrevista.notaentrevista = filas[4]
                                notalfinalenttrevista = notalfinalexamen + float(filas[4])
                                integranteentrevista.notafinal = round(notalfinalenttrevista, 0)
                                if round(notalfinalenttrevista, 0) >= notaaprobar:
                                    integranteentrevista.estado = 2
                                if round(notalfinalenttrevista, 0) < notaaprobar:
                                    integranteentrevista.estado = 3
                                integranteentrevista.save(request)
                    return JsonResponse({'result': 'ok'})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': u'Error al guardar los datos'})

        elif action == 'addcohortes':
            try:
                f = CohorteMaestriaForm(request.POST)
                maestriaadmision = MaestriasAdmision.objects.get(pk=int(encrypt(request.POST['idperiodomaestria'])))
                if f.is_valid():
                    valorexamen = 0
                    valormatricula = None
                    numerocuota = None
                    valorcuota = None
                    valortramite = 0
                    if f.cleaned_data['tienecostotramite']:
                        valortramite = f.cleaned_data['valortramite']
                    # if f.cleaned_data['tienecostoexamen']:
                    #    valorexamen = f.cleaned_data['valorexamen']
                    # if f.cleaned_data['tienecostomatricula']:
                    #     valormatricula = f.cleaned_data['valormatricula']
                    # if f.cleaned_data['tienecuota']:
                    #     numerocuota = f.cleaned_data['numerocuota']
                    #     valorcuota = f.cleaned_data['valorcuota']
                    # coordinador = Profesor.objects.get(pk=f.cleaned_data['coordinador'])
                    cohorte = CohorteMaestria.objects.filter(maestriaadmision=maestriaadmision, status=True).order_by('-id').first()
                    programamaestria = CohorteMaestria(periodoacademico=f.cleaned_data['periodoacademico'],
                                                       coordinador=f.cleaned_data['coordinador'],
                                                       maestriaadmision=maestriaadmision,
                                                       descripcion=f.cleaned_data['descripcion'],
                                                       modalidad=f.cleaned_data['modalidad'],
                                                       alias=f.cleaned_data['alias'],
                                                       numerochorte=f.cleaned_data['numerochorte'],
                                                       cupodisponible=f.cleaned_data['cupodisponible'],
                                                       # cuposlibres=f.cleaned_data['cuposlibres'],
                                                       cantidadgruposexamen=f.cleaned_data['cantidadgruposexamen'],
                                                       fechainiciocohorte=f.cleaned_data['fechainiciocohorte'],
                                                       fechafincohorte=f.cleaned_data['fechafincohorte'],
                                                       fechainicioinsp=f.cleaned_data['fechainicioinsp'],
                                                       fechafininsp=f.cleaned_data['fechafininsp'],
                                                       # fechainicioextraordinariainsp=f.cleaned_data['fechainicioextraordinariainsp'],
                                                       # fechafinextraordinariainsp=f.cleaned_data['fechafinextraordinariainsp'],
                                                       fechainiciorequisito=f.cleaned_data['finiciorequisitos'],
                                                       fechafinrequisito=f.cleaned_data['ffinrequisitos'],
                                                       fechafinrequisitobeca=f.cleaned_data['fechafinrequisitobeca'],
                                                       fechainicioexamen=f.cleaned_data['fechainicioexamen'],
                                                       fechafinexamen=f.cleaned_data['fechafinexamen'],
                                                       notaminimaexa=f.cleaned_data['notaminimaexa'],
                                                       # notamaximaexa=f.cleaned_data['notamaximaexa'],
                                                       notaminimatest=f.cleaned_data['notaminimatest'],
                                                       # notamaximatest=f.cleaned_data['notamaximatest'],
                                                       # ponderacionminimaentrevista=f.cleaned_data['ponderacionminimaentrevista'],
                                                       # ponderacionmaximaentrevista=f.cleaned_data['ponderacionmaximaentrevista'],
                                                       # tienecostoexamen=f.cleaned_data['tienecostoexamen'],
                                                       valorexamen=valorexamen,
                                                       # tienecostomatricula=f.cleaned_data['tienecostomatricula'],
                                                       # valormatricula=valormatricula,
                                                       # tienecuota=f.cleaned_data['tienecuota'],
                                                       # numerocuota=numerocuota,
                                                       # valorcuota=valorcuota,
                                                       tienecostotramite=f.cleaned_data['tienecostotramite'],
                                                       valortramite=valortramite,
                                                       activo=f.cleaned_data['activo'],
                                                       minutosrango=f.cleaned_data['minutosrango'],
                                                       cantidadgruposentrevista=f.cleaned_data[
                                                           'cantidadgruposentrevista'],
                                                       totaladmitidoscohorte=f.cleaned_data['totaladmitidoscohorte'],
                                                       )
                    programamaestria.save(request)
                    if not cohorte:
                        listadorequisitos = Requisito.objects.filter(pk__in=[2, 4, 7, 13], status=True)
                        for requisito in listadorequisitos:
                            if not RequisitosMaestria.objects.filter(cohorte=programamaestria, requisito=requisito,
                                                                     status=True).exists():
                                requisitomaestria = RequisitosMaestria(cohorte=programamaestria,
                                                                       requisito=requisito)
                                requisitomaestria.save(request)
                    else:
                        requisitosdefault = cohorte.requisitosmaestria_set.filter(status=True)
                        for requisito in requisitosdefault:
                            if not RequisitosMaestria.objects.filter(cohorte=programamaestria,
                                                                     requisito=requisito.requisito,
                                                                     status=True).exists():
                                requisitomaestria = RequisitosMaestria(cohorte=programamaestria,
                                                                       requisito=requisito.requisito)
                                requisitomaestria.save(request)
                    log(u'Adicionó programa de maestria: %s' % programamaestria, request, "add")
                    return JsonResponse({"result": False}, safe=False)
                else:
                    raise NameError('Error')
            except Exception as ex:
                msg = ex.__str__()
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos. Detalle: %s" % (msg)})

        elif action == 'editarcohorte':
            try:
                valorexamen = 0
                valormatricula = None
                numerocuota = None
                valorcuota = None
                f = CohorteMaestriaForm(request.POST)
                cohorte = CohorteMaestria.objects.get(pk=int(encrypt(request.POST['id'])))
                if f.is_valid():
                    audifechainiciocohorte = cohorte.fechainiciocohorte
                    audifechafincohorte = cohorte.fechafincohorte
                    audifechainicioinsp = cohorte.fechainicioinsp
                    audifechafininsp = cohorte.fechafininsp
                    audifechainiciorequisito = cohorte.fechainiciorequisito
                    audifechafinrequisito = cohorte.fechafinrequisito
                    cohorte.periodoacademico = f.cleaned_data['periodoacademico']
                    cohorte.descripcion = f.cleaned_data['descripcion']
                    cohorte.modalidad = f.cleaned_data['modalidad']
                    cohorte.alias = f.cleaned_data['alias']
                    cohorte.numerochorte = f.cleaned_data['numerochorte']
                    cohorte.cupodisponible = f.cleaned_data['cupodisponible']
                    # cohorte.cuposlibres = f.cleaned_data['cuposlibres']
                    cohorte.cantidadgruposexamen = f.cleaned_data['cantidadgruposexamen']
                    cohorte.cantidadgruposentrevista = f.cleaned_data['cantidadgruposentrevista']
                    cohorte.fechainiciocohorte = f.cleaned_data['fechainiciocohorte']
                    cohorte.fechafincohorte = f.cleaned_data['fechafincohorte']
                    cohorte.fechainicioinsp = f.cleaned_data['fechainicioinsp']
                    cohorte.fechafininsp = f.cleaned_data['fechafininsp']
                    # cohorte.fechainicioextraordinariainsp=f.cleaned_data['fechainicioextraordinariainsp']
                    # cohorte.fechafinextraordinariainsp=f.cleaned_data['fechafinextraordinariainsp']
                    cohorte.fechainiciorequisito = f.cleaned_data['finiciorequisitos']
                    cohorte.fechafinrequisito = f.cleaned_data['ffinrequisitos']
                    cohorte.fechafinrequisitobeca = f.cleaned_data['fechafinrequisitobeca']
                    cohorte.fechainicioexamen = f.cleaned_data['fechainicioexamen']
                    cohorte.fechafinexamen = f.cleaned_data['fechafinexamen']
                    cohorte.notaminimaexa = f.cleaned_data['notaminimaexa']
                    cohorte.notaminimatest = f.cleaned_data['notaminimatest']
                    cohorte.minutosrango = f.cleaned_data['minutosrango']
                    cohorte.tipo = f.cleaned_data['tipo']
                    cohorte.totaladmitidoscohorte = f.cleaned_data['totaladmitidoscohorte']
                    cohorte.numerocuota = numerocuota
                    cohorte.valorcuota = valorcuota
                    cohorte.activo = f.cleaned_data['activo']
                    cohorte.tienecostotramite = f.cleaned_data['tienecostotramite']
                    if f.cleaned_data['tienecostotramite']:
                        cohorte.valortramite = f.cleaned_data['valortramite']
                    else:
                        cohorte.valortramite = 0
                    idcoordinadoranterior = 0
                    idcoordinadoractual = 0
                    if cohorte.coordinador:
                        idcoordinadoranterior = cohorte.coordinador_id
                    if f.cleaned_data['coordinador']:
                        # coordinador = Persona.objects.get(pk=f.cleaned_data['coordinador'])
                        cohorte.coordinador = f.cleaned_data['coordinador']
                        idcoordinadoractual = f.cleaned_data['coordinador'].id
                    log(u'Editó cohorte: %s - cambio fecha fechainiciocohorte: %s por %s | fechafincohorte: %s por %s | fechainicioinsp %s por %s | fechafininsp %s por %s | finiciorequisitos %s por %s | ffinrequisitos %s por %s | cambio coordinador %s por coordinador %s  ' % (
                        cohorte.id, audifechainiciocohorte, f.cleaned_data['fechainiciocohorte'], audifechafincohorte,
                        f.cleaned_data['fechafincohorte'], audifechainicioinsp, f.cleaned_data['fechainicioinsp'],
                        audifechafininsp, f.cleaned_data['fechafininsp'], audifechainiciorequisito,
                        f.cleaned_data['finiciorequisitos'], audifechafinrequisito, f.cleaned_data['ffinrequisitos'],
                        idcoordinadoranterior, idcoordinadoractual), request, "edit")
                    cohorte.save(request)
                    return JsonResponse({"result": False}, safe=False)
                else:
                    raise NameError('Error')
            except Exception as ex:
                msg = ex.__str__()
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos. Detalle: %s" % (msg)})

        if action == 'addtipoclasificacionreq':
            try:
                with transaction.atomic():
                    f = TipoClasificacionReqForm(request.POST)
                    if f.is_valid():
                        if TipoClasificacionRequisito.objects.filter(descripcion=f.cleaned_data['descripcion'],
                                                                     status=True).exists():
                            transaction.set_rollback(True)
                            return JsonResponse({"result": True, "mensaje": "Ya existe el registro."}, safe=False)
                        tipo = TipoClasificacionRequisito(descripcion=f.cleaned_data['descripcion'])
                        tipo.save(request)
                        log(u'Adiciono Tipo Clasificacion Requisito: %s' % tipo, request, "add")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Complete los datos requeridos."}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'edittipoclasificacionreq':
            try:
                f = TipoClasificacionReqForm(request.POST)
                if f.is_valid():
                    tipo = TipoClasificacionRequisito.objects.get(pk=int(encrypt(request.POST['id'])))
                    tipo.descripcion = f.cleaned_data['descripcion']
                    tipo.save(request)
                    log(u'Editó Tipo Clasificacion Requisito: %s' % tipo, request, "edit")
                    return JsonResponse({"result": False}, safe=False)
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        if action == 'deletetipoclasificacionreq':
            try:
                idtipo = request.POST['id']
                tipo = TipoClasificacionRequisito.objects.get(pk=idtipo)
                log(u'Eliminó tipo de clasifición: %s' % tipo, request, "del")
                tipo.status = False
                tipo.save()
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar registro."})

        if action == 'addclasificacionrequisito':
            try:
                with transaction.atomic():
                    f = ClasificacionRequisitoForm(request.POST)
                    if f.is_valid():
                        if ClaseRequisito.objects.filter(requisito=f.cleaned_data['requisito'],
                                                         clasificacion=f.cleaned_data['clasificacion'],
                                                         status=True).exists():
                            transaction.set_rollback(True)
                            return JsonResponse({"result": True,
                                                 "mensaje": "Ya existe el requisito en %s" % f.cleaned_data[
                                                     'clasificacion']}, safe=False)
                        clasereq = ClaseRequisito(requisito=f.cleaned_data['requisito'],
                                                  clasificacion=f.cleaned_data['clasificacion'])
                        clasereq.save(request)
                        log(u'Adiciono clasificación de requisito: %s' % clasereq, request, "add")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Complete los datos requeridos."}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        if action == 'editclasificacionrequisito':
            try:
                with transaction.atomic():
                    filtro = ClaseRequisito.objects.get(pk=request.POST['id'])
                    clasificacion = TipoClasificacionRequisito.objects.get(pk=request.POST['clasificacion'])
                    if ClaseRequisito.objects.filter(requisito=filtro.requisito, clasificacion=clasificacion,
                                                     status=True).exists():
                        transaction.set_rollback(True)
                        return JsonResponse(
                            {"result": True, "mensaje": "Ya existe el requisito en %s" % (clasificacion)}, safe=False)
                    filtro.clasificacion = clasificacion
                    filtro.save(request)
                    log(u'Modificó clasificación de requisito: %s' % filtro, request, "edit")
                    return JsonResponse({"result": False}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        if action == 'edittipopersonarequisito':
            try:
                with transaction.atomic():
                    filtro = ClaseRequisito.objects.get(pk=request.POST['id'])
                    requi = Requisito.objects.get(pk=filtro.requisito.id, status=True)
                    tipoperso = TipoPersonaRequisito.objects.get(pk=request.POST['tipopersona'])
                    requi.tipopersona = tipoperso
                    requi.save(request)
                    log(u'Asignó tipo de persona a requisito: %s' % requi, request, "edit")
                    return JsonResponse({"result": False}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        if action == 'delclasificacionrequisito':
            try:
                clasereq = ClaseRequisito.objects.get(pk=request.POST['id'])
                clasereq.status = False
                clasereq.save()
                log(u'Eliminó clasificación requisito: %s' % clasereq, request, "del")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'pdfcontrol':
            try:
                data = {}
                data['cohorte'] = cohorte = CohorteMaestria.objects.get(pk=request.POST['idcohorte'])
                data['integrantes'] = Inscripcion.objects.filter(inscripcioncohorte__cohortes=cohorte,
                                                                 status=True).order_by('persona__apellido1',
                                                                                       'persona__apellido2',
                                                                                       'persona__nombres')
                cantidadcuotas = []
                for contador in range(1, cohorte.numerocuota + 1):
                    cantidadcuotas.append(contador)
                data['cantidadcuotas'] = cantidadcuotas
                return conviert_html_to_pdf(
                    'adm_admision/controlcancelaciones_pdf.html',
                    {
                        'pagesize': 'A4 landscape',
                        'data': data,
                    }
                )
            except Exception as ex:
                pass

        elif action == 'addrequisitomaestria':
            try:
                if 'archivo' in request.FILES:
                    newfile = request.FILES['archivo']
                    if newfile.size > 10485760:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Error, Tamaño de archivo Maximo permitido es de 10Mb"})
                f = AdmiRequisitosMaestriaForm(request.POST)
                if f.is_valid():
                    maestria = CohorteMaestria.objects.get(pk=int(encrypt(request.POST['id'])))
                    if not RequisitosMaestria.objects.filter(maestria=maestria,
                                                             descripcion=f.cleaned_data['descripcion'],
                                                             status=True).exists():
                        requisitomaestria = RequisitosMaestria(maestria=maestria,
                                                               descripcion=f.cleaned_data['descripcion'],
                                                               observacion=f.cleaned_data['observacion'],
                                                               activo=f.cleaned_data['activo'],
                                                               requerido=f.cleaned_data['requerido'])
                        requisitomaestria.save(request)
                        if 'archivo' in request.FILES:
                            newfile = request.FILES['archivo']
                            newfile._name = generar_nombre("formato_", newfile._name)
                            requisitomaestria.archivo = newfile
                            requisitomaestria.save(request)
                        log(u'Adicionó requisito de maestria: %s' % requisitomaestria, request, "add")
                        return JsonResponse({"result": "ok"})
                    else:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"El requisito ya existe en el programa de maestría."})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'listadepreguntas':
            try:
                idcohorte = request.POST['idcohorte']
                preguntamaestria = PreguntaMaestria.objects.values_list('pregunta_id').filter(cohortes_id=idcohorte,
                                                                                              status=True)
                listapreguntas = PreguntasPrograma.objects.filter(status=True).exclude(
                    pk__in=preguntamaestria).order_by('descripcion')
                lista = []
                for preguntas in listapreguntas:
                    datadoc = {}
                    datadoc['id'] = preguntas.id
                    datadoc['descripcion'] = preguntas.descripcion
                    lista.append(datadoc)
                return JsonResponse({'result': 'ok', 'lista': lista})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'addpreguntamaestria':
            try:
                f = AdmiPreguntasMaestriaForm(request.POST)
                if f.is_valid():
                    maestria = CohorteMaestria.objects.get(pk=int(encrypt(request.POST['id'])))
                    if not PreguntaMaestria.objects.filter(maestria=maestria, descripcion=f.cleaned_data['descripcion'],
                                                           status=True).exists():
                        preguntamaestria = PreguntaMaestria(maestria=maestria,
                                                            descripcion=f.cleaned_data['descripcion'],
                                                            activo=f.cleaned_data['activo'])
                        preguntamaestria.save(request)
                        log(u'Adicionó pregunta de maestria: %s' % preguntamaestria, request, "add")
                        return JsonResponse({"result": "ok"})
                    else:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"La pregunta ya existe en el programa de maestría."})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'aprobarrequisitoevidencia':
            try:
                detalleevidencia_aux = DetalleEvidenciaRequisitosAspirante.objects.get(pk=request.POST['idevidencia'])
                detalleevidencia = DetalleEvidenciaRequisitosAspirante(evidencia=detalleevidencia_aux.evidencia,
                                                    fecha=datetime.now().date(),
                                                    persona = persona,
                                                    estadorevision=1,
                                                    observacion = '',
                                                    estado_aprobacion  = request.POST['id_estado'],
                                                    fecha_aprobacion  = datetime.now(),
                                                    observacion_aprobacion = request.POST['id_observacion'],)
                # detalleevidencia.observacion_aprobacion = request.POST['id_observacion']
                # detalleevidencia.persona = persona
                # detalleevidencia.estado_aprobacion = request.POST['id_estado']
                # detalleevidencia.fecha_aprobacion = datetime.now()
                detalleevidencia.save(request)

                log(u'Creó observacion evidencia: %s' % (detalleevidencia), request, "add")

                # enviar mail cuando aprueban requisitos posgrado
                # if detalleevidencia.estado_aprobacion == '2':
                #     send_html_mail("APROBADA EVIDENCIA - UNEMI.", "emails/registroaprobadoevidencia.html",
                #                    {'sistema': u'ADMISIÓN - UNEMI', 'fecha': datetime.now().date(),
                #                     'hora': datetime.now().time(), 't': miinstitucion(),
                #                     'requisito': detalleevidencia.evidencia.requisitos.requisito.nombre,
                #                     'observacion': detalleevidencia.observacion_aprobacion},
                #                    detalleevidencia.evidencia.inscripcioncohorte.inscripcionaspirante.persona.emailpersonal(),
                #                    [],
                #                    cuenta=variable_valor('CUENTAS_CORREOS')[16])

                # enviar mail cuando le rechazan requisitos posgrado
                if detalleevidencia.estado_aprobacion == '3':
                    asunto = u"REQUISITO DE ADMISIÓN RECHAZADO"
                    observacion = f'Se le comunica que la Secretaría Técnica de Posgrado ha rechazado el requisito {detalleevidencia.evidencia.requisitos.requisito.nombre}, con la siguiente observación {detalleevidencia.observacion_aprobacion}. Por favor, comunicarse con el postulante {detalleevidencia.evidencia.inscripcioncohorte.inscripcionaspirante.persona.nombre_completo_inverso()} para subir el requisito correctamente. Puede dar clic en la URL para ser redirigido directamente al postulante  que presenta dicha observación.'
                    para = detalleevidencia.evidencia.inscripcioncohorte.asesor.persona
                    perfiu = detalleevidencia.evidencia.inscripcioncohorte.asesor.perfil_administrativo()
                    notificacion3(asunto, observacion, para, None, '/comercial?s=' + detalleevidencia.evidencia.inscripcioncohorte.inscripcionaspirante.persona.cedula, detalleevidencia.pk, 1,
                                 'sga', DetalleEvidenciaRequisitosAspirante, perfiu, request)

                    inscrito = InscripcionCohorte.objects.get(status=True, id=detalleevidencia.evidencia.inscripcioncohorte.id)
                    inscrito.preaprobado = False
                    inscrito.todosubido = False
                    inscrito.tienerechazo = True

                    deta = DetallePreAprobacionPostulante(inscripcion=inscrito,
                                                          preaprobado=inscrito.preaprobado)
                    deta.save(request)

                    if inscrito.estado_aprobador == 2:
                        inscrito.estado_aprobador = 1
                    inscrito.save(request)

                    log(u'Postulante con evidencia rechazada: %s' % (inscrito), request, "edit")

                    # send_html_mail("RECHAZADA EVIDENCIA - UNEMI.", "emails/registrorechazadoevidencia.html",
                    #                {'sistema': u'ADMISIÓN - UNEMI', 'fecha': datetime.now().date(),
                    #                 'hora': datetime.now().time(), 't': miinstitucion(),
                    #                 'requisito': detalleevidencia.evidencia.requisitos.requisito.nombre,
                    #                 'observacion': detalleevidencia.observacion_aprobacion},
                    #                detalleevidencia.evidencia.inscripcioncohorte.inscripcionaspirante.persona.emailpersonal(),
                    #                [],
                    #                cuenta=variable_valor('CUENTAS_CORREOS')[16])

                cohorte = detalleevidencia.evidencia.inscripcioncohorte.cohortes
                inscripcioncohorte = detalleevidencia.evidencia.inscripcioncohorte
                # verificar si ya esta todo aprobado para enviar correo de aprobacion

                claserequisitoadmision = ClaseRequisito.objects.values_list('requisito__id').filter(clasificacion=1,
                                                                                                    status=True)
                requisitosadmision = cohorte.requisitosmaestria_set.filter(requisito__in=claserequisitoadmision,
                                                                           obligatorio=True, status=True).order_by('id')
                bandera = 0
                for re in requisitosadmision:
                    ingresoevidencias = re.detalle_requisitosmaestriacohorte(inscripcioncohorte)
                    if ingresoevidencias:
                        if not ingresoevidencias.ultima_evidencia().estado_aprobacion == 2:
                            bandera = 1
                    else:
                        bandera = 1
                if bandera == 0:
                    # inscripcioncoohorte = InscripcionCohorte.objects.get(pk=request.POST['idinscripcioncohorte'])
                    inscripcioncohorte.fecha_emailevidencia = datetime.now()
                    inscripcioncohorte.estado_emailevidencia = 2
                    inscripcioncohorte.persona_emailevidencia = persona
                    inscripcioncohorte.fecha_aprobador = datetime.now()
                    inscripcioncohorte.estado_aprobador = 2
                    inscripcioncohorte.persona_aprobador = persona
                    inscripcioncohorte.tienerechazo = False
                    inscripcioncohorte.save(request)

                    asunto = u"POSTULANTE ADMITIDO"
                    observacion = f'Se le comunica que el postulante {inscripcioncohorte.inscripcionaspirante.persona} con cédula {inscripcioncohorte.inscripcionaspirante.persona.cedula} ha sido admitido en el programa de {inscripcioncohorte.cohortes.maestriaadmision.descripcion}. Por favor, brindar el respectivo seguimiento.'
                    para = inscripcioncohorte.asesor.persona
                    perfiu = inscripcioncohorte.asesor.perfil_administrativo()

                    notificacion3(asunto, observacion, para, None,
                                  '/comercial?s=' + inscripcioncohorte.inscripcionaspirante.persona.cedula,
                                  inscripcioncohorte.pk, 1,
                                  'sga', InscripcionCohorte, perfiu, request)

                    log(u'Envio email aprobacion, masivo: %s' % (inscripcioncohorte), request, "add")

                    banneradjunto = archivoadjunto = ''
                    if FormatoCarreraIpec.objects.filter(carrera_id=inscripcioncohorte.cohortes.maestriaadmision.carrera.id, status=True):
                        formatocorreo = FormatoCarreraIpec.objects.filter(carrera_id=inscripcioncohorte.cohortes.maestriaadmision.carrera.id, status=True)[0]
                        archivoadjunto = formatocorreo.archivo
                        banneradjunto = formatocorreo.banner

                    mencion = ''
                    if inscripcioncohorte.cohortes.maestriaadmision.carrera.malla().tiene_itinerario_malla_especialidad() == True:
                        iti = ItinerarioMallaEspecilidad.objects.get(status=True, malla=inscripcioncohorte.cohortes.maestriaadmision.carrera.malla(),
                                                                     itinerario=inscripcioncohorte.itinerario)
                        mencion = iti.nombre

                    send_html_mail("Aprobado Admision-UNEMI.", "emails/registroaprobadomasivo.html",
                                   {'sistema': u'Admision - UNEMI', 'fecha': datetime.now().date(),
                                    'hora': datetime.now().time(), 't': miinstitucion(), 'formato': banneradjunto,
                                    'inscripcioncohorte': inscripcioncohorte,
                                    'nomcarrera': inscripcioncohorte.cohortes.maestriaadmision.carrera.nombre[11:] if inscripcioncohorte.cohortes.maestriaadmision.carrera.nombre[:6] == 'MAESTR' else inscripcioncohorte.cohortes.maestriaadmision.carrera.nombre,
                                    'mencion':mencion},
                                   inscripcioncohorte.inscripcionaspirante.persona.emailpersonal(), [],
                                   cuenta=variable_valor('CUENTAS_CORREOS')[16])
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al aprobar."})

        elif action == 'aprobarcontratoevidencia':
            try:
                espagare = True if 'espagare' in request.POST else False
                detalleevidencia = DetalleAprobacionContrato(contrato_id=request.POST['idevidencia'], espagare=espagare)
                detalleevidencia.save(request)
                detalleevidencia.observacion = request.POST['id_observacion']
                detalleevidencia.persona = persona
                detalleevidencia.estado_aprobacion = request.POST['id_estado']
                detalleevidencia.fecha_aprobacion = datetime.now()
                detalleevidencia.save(request)
                log(u'Actualizó observacion evidencia: %s' % (detalleevidencia), request, "add")

                # enviar mail cuando aprueban contrato posgrado
                if detalleevidencia.estado_aprobacion == '2':
                    send_html_mail("APROBADA EVIDENCIA - UNEMI.", "emails/registroaprobadoevidencia.html",
                                   {'sistema': u'ADMISIÓN - UNEMI', 'fecha': datetime.now().date(),
                                    'hora': datetime.now().time(), 't': miinstitucion(),
                                    'requisito': 'Contrato de pago - %s' % detalleevidencia.contrato.formapago,
                                    'observacion': detalleevidencia.observacion},
                                   detalleevidencia.persona.emailpersonal(), [],
                                   cuenta=variable_valor('CUENTAS_CORREOS')[16])

                # enviar mail cuando le rechazan requisitos posgrado
                if detalleevidencia.estado_aprobacion == '3':
                    send_html_mail("RECHAZADA EVIDENCIA - UNEMI.", "emails/registrorechazadoevidencia.html",
                                   {'sistema': u'ADMISIÓN - UNEMI', 'fecha': datetime.now().date(),
                                    'hora': datetime.now().time(), 't': miinstitucion(),
                                    'requisito': 'Contrato de pago - %s' % detalleevidencia.contrato.formapago,
                                    'observacion': detalleevidencia.observacion},
                                   detalleevidencia.persona.emailpersonal(), [],
                                   cuenta=variable_valor('CUENTAS_CORREOS')[16])

                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al aprobar o rechazar evidencia."})

        if action == 'pdfcontratopagoprograma':
            try:
                numcontrato = 0
                idins = request.POST['idins']
                admitido = InscripcionCohorte.objects.get(status=True, pk=int(idins))
                registro = Contrato.objects.filter(status=True, inscripcion__id=idins, inscripcion__status=True).last()
                secuenciacp = secuencia_contratopagare(request, datetime.now().year)
                if not registro or not registro.numerocontrato:
                    secuenciacp.secuenciacontrato += 1
                    secuenciacp.save(request)
                    if Contrato.objects.filter(status=True, numerocontrato=secuenciacp.secuenciacontrato, fechacontrato__year=secuenciacp.anioejercicio.anioejercicio).exists():
                        return JsonResponse({"result": "bad", "mensaje": u"Error al generar el contrato, intente nuevamente"})
                    else:
                        if Contrato.objects.filter(status=True,inscripcion_id=idins).exists():
                            ct = Contrato.objects.filter(inscripcion_id=idins, status=True,).last()
                            ct.numerocontrato = secuenciacp.secuenciacontrato
                            ct.save(request)
                            log(u'Editó número contrato: %s' % (ct), request, "add")
                        else:
                            ct = Contrato(inscripcion_id=idins, numerocontrato=secuenciacp.secuenciacontrato)
                            ct.save(request)
                            log(u'Adicionó numero contrato: %s' % (ct), request, "add")
                        numcontrato = ct.numerocontrato
                else:
                    if Contrato.objects.get(status=True, inscripcion__id=idins, inscripcion__status=True).numerocontrato:
                        numcontrato = Contrato.objects.get(status=True, inscripcion__id=idins, inscripcion__status=True).numerocontrato

                cont = Contrato.objects.get(status=True, inscripcion__id=idins)

                tipo = 'pdf'
                paRequest = {
                    'idins': admitido.id,
                    'numcontrato': numcontrato
                }

                reporte = None
                if admitido.formapagopac.id == 1:
                    reporte = Reporte.objects.get(id=663)
                elif admitido.formapagopac.id == 2:
                    reporte = Reporte.objects.get(id=664)

                d = run_report_v1(reporte=reporte, tipo=tipo, paRequest=paRequest, request=request)

                if not d['isSuccess']:
                    raise NameError(d['mensaje'])
                else:
                    url_archivo = (SITE_STORAGE + d['data']['reportfile']).replace('\\', '/')
                    url_archivo = (url_archivo).replace('//', '/')
                    _name = generar_nombre(f'contrato_{request.user.username}_{idins}_','descargado')
                    folder = os.path.join(SITE_STORAGE, 'media', 'archivodescargado', '')
                    if not os.path.exists(folder):
                        os.makedirs(folder)
                    folder_save = os.path.join('archivodescargado', '').replace('\\', '/')
                    url_file_generado = f'{folder_save}{_name}.pdf'
                    ruta_creacion = SITE_STORAGE
                    ruta_creacion = ruta_creacion.replace('\\', '/')
                    shutil.copy(url_archivo, ruta_creacion + '/media/' + url_file_generado)

                    cont.archivodescargado = url_file_generado
                    cont.save(request)
                    # return JsonResponse({"result": "ok", 'url': cont.download_descargado()})
                    return JsonResponse({"result": "ok", 'url': d['data']['reportfile']})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al descargar.%s"%(ex)})

        if action == 'cargarcontratopago':
            try:
                hayarchivos = False
                if 'archivo' in request.FILES:
                    hayarchivos = True
                    descripcionarchivo = 'Contrato de pago'
                    resp = validar_archivo(descripcionarchivo, request.FILES['archivo'], ['pdf'], '4MB')
                    if resp['estado'] != "OK":
                        return JsonResponse({"result": "bad", "mensaje": resp["mensaje"]})
                else:
                    return JsonResponse({"result": "bad", "mensaje": u"Debe seleccionar un archivo (Contrato)."})

                f = ContratoPagoMaestriaForm(request.POST, request.FILES)
                if f.is_valid():
                    if hayarchivos:
                        contratopago = request.FILES['archivo']
                        contratopago._name = generar_nombre("contratopago", contratopago._name)
                        insc = InscripcionCohorte.objects.get(pk=request.POST['id'])
                        if Contrato.objects.filter(status=True, inscripcion=insc, inscripcion__status=True).exists():
                            contrato = Contrato.objects.get(status=True, inscripcion=insc, inscripcion__status=True)
                            contrato.fechacontrato = hoy
                            contrato.formapago_id = int(request.POST['fpago'])
                            contrato.estado = 1
                            contrato.archivocontrato = contratopago
                            contrato.observacion = request.POST['observacion']
                            contrato.save(request)
                            log(u'Adicionó Contrato de : %s' % (insc), request, "edit")

                            detalle = DetalleAprobacionContrato(contrato=contrato,
                                                                estado_aprobacion=1,
                                                                observacion='')
                            detalle.save(request)
                            return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar contrato."})

        elif action == 'generartablaamortizacion':
            try:
                inscripcion = InscripcionCohorte.objects.get(pk=request.POST['idaspirante'])
                infofinancieramae = InfraestructuraEquipamientoInformacionPac.objects.filter(
                    programapac__carrera=inscripcion.cohortes.maestriaadmision.carrera).last()
                if infofinancieramae.valorarancel:
                    contrato = inscripcion.contrato_set.last()
                    valorarancelmae = infofinancieramae.valorarancel
                    maxnumcuota = inscripcion.numcuotaspago
                    porcentajemae = infofinancieramae.porcentajeminpagomatricula
                    nombremae = inscripcion.cohortes
                    fechainiciopago = hoy
                    fechafinpago = hoy + timedelta(days=3)
                    # Generacion de la primera cuota aplicando porcentaje de la cuota inicial
                    valordescontado = Decimal(null_to_decimal((valorarancelmae * porcentajemae) / 100, 2)).quantize(
                        Decimal('.01'))
                    tamortizacioninicial = TablaAmortizacion(
                        contrato=contrato,
                        numerocuota=1,
                        nombre='TA' + ' - ' + str(nombremae),
                        valor=valordescontado,
                        fechainiciopago=fechainiciopago,
                        fechafinpago=fechafinpago)
                    tamortizacioninicial.save(request)

                    # Generacion de cuotas
                    valorrestante = Decimal(null_to_decimal(valorarancelmae)).quantize(Decimal('.01')) - valordescontado
                    valor = Decimal(null_to_decimal(valorrestante / (maxnumcuota - 1))).quantize(Decimal('.01'))
                    for i in range(2, maxnumcuota):
                        fechainiciopago = fechafinpago
                        fechafinpago = fechainiciopago + timedelta(days=3)
                        tamortizacion = TablaAmortizacion(
                            contrato=contrato,
                            numerocuota=i,
                            nombre='TA' + ' - ' + str(nombremae),
                            valor=valor,
                            fechainiciopago=fechainiciopago,
                            fechafinpago=fechafinpago)
                        tamortizacion.save(request)

                    # Generacion de cuota final para evitar perder decimales($)
                    valorfinal = Decimal(null_to_decimal(valorarancelmae)).quantize(Decimal('.01')) - (
                                valor * (maxnumcuota - 2) + valordescontado)
                    fechainiciopago = fechafinpago
                    fechafinpago = fechainiciopago + timedelta(days=3)
                    tamortizacion = TablaAmortizacion(
                        contrato=contrato,
                        numerocuota=maxnumcuota,
                        nombre='TA' + ' - ' + str(nombremae),
                        valor=valorfinal,
                        fechainiciopago=fechainiciopago,
                        fechafinpago=fechafinpago)
                    tamortizacion.save(request)

                log(u'Genero tabla de amortización: %s programa de %s' % (
                inscripcion.inscripcionaspirante, inscripcion.cohortes.maestriaadmision.descripcion), request, "add")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})


        elif action == 'aprobarnotificacion':
            try:
                browser = request.POST['navegador']
                ops = request.POST['os']
                cookies = request.POST['cookies']
                screensize = request.POST['screensize']
                inscripcioncoohorte = InscripcionCohorte.objects.get(pk=request.POST['idinscripcioncohorte'])
                inscripcioncoohorte.fecha_emailevidencia = datetime.now()
                inscripcioncoohorte.estado_emailevidencia = 2
                inscripcioncoohorte.persona_emailevidencia = persona
                inscripcioncoohorte.save(request)
                log(u'Envio email aprobacion: %s' % (inscripcioncoohorte), request, "add")
                send_html_mail("Aprobado Admision-UNEMI.", "emails/registroaprobado.html",
                               {'sistema': u'Admision - UNEMI', 'fecha': datetime.now().date(),
                                'inscripcioncoohorte': inscripcioncoohorte,
                                'hora': datetime.now().time(), 'bs': browser, 'os': ops, 'cookies': cookies,
                                'screensize': screensize, 't': miinstitucion()},
                               inscripcioncoohorte.inscripcionaspirante.persona.emailpersonal(), [],
                               cuenta=variable_valor('CUENTAS_CORREOS')[16])
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al aprobar."})

        elif action == 'cambiocohorte':
            try:
                inscripcioncoohorte = InscripcionCohorte.objects.get(pk=request.POST['idinscripcioncohorteotracohorte'])
                log(u'%s - Cambio cohorte: %s a la cohorte numero %s' % (
                    inscripcioncoohorte.cohortes_id, inscripcioncoohorte, request.POST['id_cohorte']), request, "add")
                listarequisitos = EvidenciaRequisitosAspirante.objects.filter(inscripcioncohorte=inscripcioncoohorte)
                for lis in listarequisitos:
                    if RequisitosMaestria.objects.filter(requisito=lis.requisitos.requisito,
                                                         cohorte_id=request.POST['id_cohorte'], status=True):
                        requi = RequisitosMaestria.objects.filter(requisito=lis.requisitos.requisito,
                                                                  cohorte_id=request.POST['id_cohorte'], status=True)[0]
                        lis.requisitos = requi
                        lis.save()

                if Rubro.objects.filter(inscripcion=inscripcioncoohorte, status=True).exists():
                    aspiranteconrubro = Rubro.objects.get(inscripcion=inscripcioncoohorte, status=True)
                    rubro = Rubro.objects.get(id=aspiranteconrubro.id)
                    chorte = CohorteMaestria.objects.get(id=request.POST['id_cohorte'], status=True)

                    if aspiranteconrubro.admisionposgradotipo == 2:
                        tiporubroarancel = TipoOtroRubro.objects.get(pk=2845)
                        rubro.nombre = tiporubroarancel.nombre + ' - ' + chorte.maestriaadmision.descripcion + ' - ' + chorte.descripcion
                    elif aspiranteconrubro.admisionposgradotipo == 3:
                        rubro.nombre = chorte.maestriaadmision.descripcion + ' - ' + chorte.descripcion

                    rubro.cohortemaestria = chorte
                    rubro.save(request)

                inscripcioncoohorte.cohortes_id = request.POST['id_cohorte']
                inscripcioncoohorte.save(request)

                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al aprobar, requisitos no aprobados."})

        elif action == 'cambiocohortemasiva':
            try:

                cohorte = CohorteMaestria.objects.get(pk=int(request.POST['id_cohortemasivamaestria']))
                aspirantes = InscripcionCohorte.objects.filter(estado_aprobador=1, cohortes=cohorte,
                                                               status=True).order_by(
                    'inscripcionaspirante__persona__apellido1', 'inscripcionaspirante__persona__apellido2',
                    'inscripcionaspirante__persona__nombres')

                for aspirante in aspirantes:
                    inscripcioncoohorte = InscripcionCohorte.objects.get(pk=aspirante.id)
                    log(u'%s - Cambio cohorte: %s a la cohorte numero %s' % (
                        inscripcioncoohorte.cohortes_id, inscripcioncoohorte, request.POST['id_cohortemasiva']),
                        request,
                        "add")
                    listarequisitos = EvidenciaRequisitosAspirante.objects.filter(
                        inscripcioncohorte=inscripcioncoohorte)
                    for lis in listarequisitos:
                        if RequisitosMaestria.objects.filter(requisito=lis.requisitos.requisito,
                                                             cohorte_id=request.POST['id_cohortemasiva'], status=True):
                            requi = RequisitosMaestria.objects.filter(requisito=lis.requisitos.requisito,
                                                                      cohorte_id=request.POST['id_cohortemasiva'],
                                                                      status=True)[0]
                            lis.requisitos = requi
                            lis.save()

                    aspiranteconrubro = Rubro.objects.filter(inscripcion=inscripcioncoohorte,
                                                             admisionposgradotipo__in=[2, 3],
                                                             cohortemaestria_id=inscripcioncoohorte.cohortes_id,
                                                             status=True)
                    for crubro in aspiranteconrubro:
                        crubro.cohortemaestria = request.POST['id_cohortemasiva']
                        crubro.save()

                    inscripcioncoohorte.cohortes_id = request.POST['id_cohortemasiva']
                    inscripcioncoohorte.save(request)

                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad",
                                     "mensaje": u"Error al aprobar. Verficar cohorte tengo todos los campos correcto."})

        elif action == 'cohortemaestria_selecionados':
            try:
                datos = json.loads(request.POST['lista_items2'])
                for dato in datos:
                    inscripcioncoohorte = InscripcionCohorte.objects.get(pk=dato['id'])
                    log(u'%s - Cambio cohorte: %s a la cohorte numero %s' % (
                        inscripcioncoohorte.cohortes_id, inscripcioncoohorte, request.POST['id_cohorteselecionados']),
                        request,
                        "add")
                    listarequisitos = EvidenciaRequisitosAspirante.objects.filter(
                        inscripcioncohorte=inscripcioncoohorte)
                    for lis in listarequisitos:
                        if RequisitosMaestria.objects.filter(requisito=lis.requisitos.requisito,
                                                             cohorte_id=request.POST['id_cohorteselecionados'],
                                                             status=True):
                            requi = RequisitosMaestria.objects.filter(requisito=lis.requisitos.requisito,
                                                                      cohorte_id=request.POST['id_cohorteselecionados'],
                                                                      status=True)[0]
                            lis.requisitos = requi
                            lis.save()

                    aspiranteconrubro = Rubro.objects.filter(inscripcion=inscripcioncoohorte,
                                                             admisionposgradotipo__in=[2, 3],
                                                             cohortemaestria_id=inscripcioncoohorte.cohortes_id,
                                                             status=True)
                    for crubro in aspiranteconrubro:
                        crubro.cohortemaestria = request.POST['id_cohorteselecionados']
                        crubro.save()

                    inscripcioncoohorte.cohortes_id = request.POST['id_cohorteselecionados']
                    inscripcioncoohorte.save(request)

                return JsonResponse({"result": "ok"})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": "%s" % ex})

        elif action == 'aprobarnotificacionadmitido':
            try:
                browser = request.POST['navegador']
                ops = request.POST['os']
                cookies = request.POST['cookies']
                screensize = request.POST['screensize']
                inscripcioncoohorte = IntegranteGrupoEntrevitaMsc.objects.get(pk=request.POST['idinscripcioncohorte'])
                inscripcioncoohorte.fecha_emailadmitido = datetime.now()
                inscripcioncoohorte.estado_emailadmitido = 2
                inscripcioncoohorte.persona_emailadmitido = persona
                inscripcioncoohorte.save(request)
                log(u'Envio email aprobacion: %s' % (inscripcioncoohorte), request, "add")
                send_html_mail("Aprobado Admision-UNEMI.", "emails/notificacionadmitido.html",
                               {'sistema': u'Admision - UNEMI', 'fecha': datetime.now().date(),
                                'hora': datetime.now().time(), 'bs': browser, 'os': ops, 'cookies': cookies,
                                'screensize': screensize, 't': miinstitucion()},
                               inscripcioncoohorte.inscripcion.inscripcionaspirante.persona.emailpersonal(), [],
                               cuenta=variable_valor('CUENTAS_CORREOS')[16])
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al aprobar."})

        elif action == 'admitirotracohorte':
            try:
                browser = request.POST['navegador']
                ops = request.POST['os']
                cookies = request.POST['cookies']
                screensize = request.POST['screensize']
                cadena = request.POST['listascodigos'].split(',')
                idlistacohorte = CohorteMaestria.objects.get(pk=request.POST['idlistacohorte'])
                for elemento in cadena:
                    inscripcioncoohorte = IntegranteGrupoEntrevitaMsc.objects.get(pk=elemento)
                    inscripcioncoohorte.cohorteadmitidasinproceso = idlistacohorte
                    inscripcioncoohorte.save(request)
                    log(u'Envio email admitido a otra cohorte: %s a la cohorte %s' % (
                        inscripcioncoohorte, idlistacohorte), request, "add")
                    send_html_mail("Aprobado Admision-UNEMI.", "emails/notificacionadmitidootracohorte.html",
                                   {'sistema': u'Admision - UNEMI', 'fecha': datetime.now().date(),
                                    'hora': datetime.now().time(), 'bs': browser, 'os': ops, 'cookies': cookies,
                                    'screensize': screensize, 't': miinstitucion()},
                                   inscripcioncoohorte.inscripcion.inscripcionaspirante.persona.emailpersonal(), [],
                                   cuenta=variable_valor('CUENTAS_CORREOS')[16])
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al aprobar."})

        elif action == 'listadoaprobados':
            try:
                tipoestado = int(request.POST['tipoestado'])
                cohorte = CohorteMaestria.objects.get(pk=int(encrypt(request.POST['idcohorte'])))
                inscripcioncoohorte = cohorte.inscripcioncohorte_set.values('id', 'inscripcionaspirante__persona_id',
                                                                            'inscripcionaspirante__persona__apellido1',
                                                                            'inscripcionaspirante__persona__apellido2',
                                                                            'inscripcionaspirante__persona__nombres').filter(
                    estado_aprobador=tipoestado, estado_emailevidencia=1).order_by(
                    'inscripcionaspirante__persona__apellido1', 'inscripcionaspirante__persona__apellido2',
                    'inscripcionaspirante__persona__nombres')
                return JsonResponse({"result": "ok", "cantidad": len(inscripcioncoohorte),
                                     "inscripcioncoohorte": list(inscripcioncoohorte)})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al aprobar."})

        elif action == 'listadonotificargrupoentrevista':
            try:
                grupoexamen = GrupoExamenMsc.objects.get(pk=request.POST['idgrupoexamen'])
                grupoexamen.estado_emailentrevista = 2
                grupoexamen.persona_emailentrevista = persona
                grupoexamen.save(request)
                listagrupoexamenes = grupoexamen.integrantegrupoexamenmsc_set.values('id',
                                                                                     'inscripcion__inscripcionaspirante__persona_id',
                                                                                     'inscripcion__inscripcionaspirante__persona__apellido1',
                                                                                     'inscripcion__inscripcionaspirante__persona__apellido2',
                                                                                     'inscripcion__inscripcionaspirante__persona__nombres').filter(
                    status=True).order_by('inscripcion__inscripcionaspirante__persona__apellido1',
                                          'inscripcion__inscripcionaspirante__persona__apellido2',
                                          'inscripcion__inscripcionaspirante__persona__nombres')
                return JsonResponse({"result": "ok", "cantidad": len(listagrupoexamenes),
                                     "listagrupoexamenes": list(listagrupoexamenes)})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al aprobar."})

        elif action == 'listadoexportargrupomoodle':
            try:
                grupoexamen = GrupoExamenMsc.objects.get(pk=request.POST['idgrupoexamenmoodle'])
                listadosmoodle = []
                if grupoexamen.idgrupomoodle > 0:
                    listadocodigo = grupoexamen.integrantegrupoexamenmsc_set.values_list(
                        'inscripcion__inscripcionaspirante__persona__usuario__username', flat=True).filter(status=True)
                    if listadocodigo.__len__() > 0:
                        cursorpos = connections['moodle_pos'].cursor()
                        sql = """SELECT DISTINCT  ARRAY_TO_STRING(array_agg(us1.id),',')                                      
                                 FROM mooc_role_assignments asi
                                INNER JOIN MOOC_CONTEXT CON ON asi.CONTEXTID=CON.ID
                                INNER JOIN mooc_user us1 ON us1.id=asi.userid
                                AND ASI.ROLEID=%s
                                AND CON.INSTANCEID=%s
                                AND us1.username in %s""" % (10, grupoexamen.idgrupomoodle, tuple(listadocodigo))
                        cursorpos.execute(sql)
                        row = cursorpos.fetchall()
                        if grupoexamen.idgrupomoodle:
                            if row[0][0]:
                                listadosmoodle = row[0][0].split(",")
                        cursorpos.close()
                listagrupoexamenes = grupoexamen.integrantegrupoexamenmsc_set.values('id',
                                                                                     'inscripcion__inscripcionaspirante__persona_id',
                                                                                     'inscripcion__inscripcionaspirante__persona__apellido1',
                                                                                     'inscripcion__inscripcionaspirante__persona__apellido2',
                                                                                     'inscripcion__inscripcionaspirante__persona__nombres').filter(
                    inscripcion__status=True, status=True).exclude(
                    inscripcion__inscripcionaspirante__persona__idusermoodleposgrado__in=listadosmoodle).order_by(
                    'inscripcion__inscripcionaspirante__persona__apellido1',
                    'inscripcion__inscripcionaspirante__persona__apellido2',
                    'inscripcion__inscripcionaspirante__persona__nombres')

                return JsonResponse({"result": "ok", "cantidad": len(listagrupoexamenes),
                                     "listagrupoexamenes": list(listagrupoexamenes)})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al aprobar."})

        elif action == 'exportarinscrito':
            try:
                contador = int(request.POST['contador'])
                inscritoexamen = request.POST['inscritoexamen']
                grupo = GrupoExamenMsc.objects.get(status=True, pk=request.POST['idgrupoexamenmoodle'])
                codidointegrante = grupo.integrantegrupoexamenmsc_set.get(pk=inscritoexamen, status=True)
                codidointegrante.encursomoodle = True
                codidointegrante.save(request)
                grupo.crear_grupo_moodle(inscritoexamen, contador)
                time.sleep(3)
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'enviarnotificacionrevision':
            try:
                browser = request.POST['navegador']
                ops = request.POST['os']
                cookies = request.POST['cookies']
                screensize = request.POST['screensize']
                tipoestado = int(request.POST['tipoestado'])
                cohorte = CohorteMaestria.objects.get(pk=int(encrypt(request.POST['idcohorte'])))
                inscripcioncoohorte = cohorte.inscripcioncohorte_set.filter(estado_aprobador=tipoestado,
                                                                            estado_emailevidencia=1)[:40]
                for lista in inscripcioncoohorte:
                    lista.fecha_emailevidencia = datetime.now()
                    lista.estado_emailevidencia = tipoestado
                    lista.persona_emailevidencia = persona
                    lista.save(request)
                    log(u'Envio email aprobacion: %s' % (inscripcioncoohorte), request, "add")
                    if tipoestado == 2:
                        send_html_mail("Aprobado Admision-UNEMI.", "emails/registroaprobado.html",
                                       {'sistema': u'Admision - UNEMI', 'fecha': datetime.now().date(),
                                        'hora': datetime.now().time(), 'bs': browser, 'os': ops, 'cookies': cookies,
                                        'screensize': screensize, 't': miinstitucion()},
                                       lista.inscripcionaspirante.persona.emailpersonal(), [],
                                       cuenta=variable_valor('CUENTAS_CORREOS')[16])
                    if tipoestado == 3:
                        send_html_mail("Rechazado Admision-UNEMI.", "emails/registrorechazado.html",
                                       {'sistema': u'Admision - UNEMI', 'fecha': datetime.now().date(),
                                        'hora': datetime.now().time(), 'bs': browser, 'os': ops, 'cookies': cookies,
                                        'screensize': screensize, 't': miinstitucion()},
                                       lista.inscripcionaspirante.persona.emailpersonal(), [],
                                       cuenta=variable_valor('CUENTAS_CORREOS')[16])
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al aprobar."})

        elif action == 'enviarnotificacionrevisionindividual':
            try:
                browser = request.POST['navegador']
                ops = request.POST['os']
                cookies = request.POST['cookies']
                screensize = request.POST['screensize']
                tipoestado = int(request.POST['tipoestado'])
                inscripcioncoohorte = InscripcionCohorte.objects.get(pk=request.POST['inscritocohorte'])
                inscripcioncoohorte.fecha_emailevidencia = datetime.now()
                inscripcioncoohorte.estado_emailevidencia = tipoestado
                inscripcioncoohorte.persona_emailevidencia = persona
                inscripcioncoohorte.save(request)
                arregloemail = [23, 24, 25, 26, 27, 28]
                emailaleatorio = random.choice(arregloemail)
                log(u'Envio email aprobacion: %s' % (inscripcioncoohorte), request, "add")
                if tipoestado == 2:
                    send_html_mail("Aprobado Admision-UNEMI.", "emails/registroaprobado.html",
                                   {'sistema': u'Admision - UNEMI', 'fecha': datetime.now().date(),
                                    'hora': datetime.now().time(), 'bs': browser, 'os': ops, 'cookies': cookies,
                                    'screensize': screensize, 't': miinstitucion()},
                                   inscripcioncoohorte.inscripcionaspirante.persona.emailpersonal(), [],
                                   cuenta=variable_valor('CUENTAS_CORREOS')[emailaleatorio])
                if tipoestado == 3:
                    send_html_mail("Rechazado Admision-UNEMI.", "emails/registrorechazado.html",
                                   {'sistema': u'Admision - UNEMI', 'fecha': datetime.now().date(),
                                    'hora': datetime.now().time(), 'bs': browser, 'os': ops, 'cookies': cookies,
                                    'screensize': screensize, 't': miinstitucion()},
                                   inscripcioncoohorte.inscripcionaspirante.persona.emailpersonal(), [],
                                   cuenta=variable_valor('CUENTAS_CORREOS')[emailaleatorio])
                time.sleep(5)
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al aprobar."})

        elif action == 'permisosubirevidencia':
            try:
                browser = request.POST['navegador']
                ops = request.POST['os']
                cookies = request.POST['cookies']
                screensize = request.POST['screensize']
                inscripcioncoohorte = InscripcionCohorte.objects.get(pk=request.POST['idinscripcionsubir'])
                inscripcioncoohorte.fecha_permisosubir = datetime.now()
                inscripcioncoohorte.estado_aprobador = 1
                inscripcioncoohorte.estado_emailevidencia = 1
                inscripcioncoohorte.persona_permisosubir = persona
                inscripcioncoohorte.save(request)
                log(u'Envio email permiso subir evidencia: %s' % (inscripcioncoohorte), request, "add")
                send_html_mail("Subir evidencias Admision-UNEMI.", "emails/permisosubirevidencia.html",
                               {'sistema': u'Admision - UNEMI', 'fecha': datetime.now().date(),
                                'hora': datetime.now().time(), 'bs': browser, 'os': ops, 'cookies': cookies,
                                'screensize': screensize, 't': miinstitucion()},
                               inscripcioncoohorte.inscripcionaspirante.persona.emailpersonal(), [],
                               cuenta=variable_valor('CUENTAS_CORREOS')[16])
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al aprobar."})

        elif action == 'aprobarestado':
            try:
                inscripcioncoohorte = InscripcionCohorte.objects.get(pk=request.POST['idinscripcioncohorteestado'])
                inscripcioncoohorte.fecha_aprobador = datetime.now()
                inscripcioncoohorte.estado_aprobador = request.POST['estadoaprobador']
                inscripcioncoohorte.persona_aprobador = persona
                inscripcioncoohorte.save(request)

                if request.POST['estadoaprobador'] == '2':
                    if inscripcioncoohorte.cohortes.tienecostoexamen:
                        tiporubroarancel = TipoOtroRubro.objects.get(pk=3077)
                        rubro = Rubro(tipo=tiporubroarancel,
                                      persona=inscripcioncoohorte.inscripcionaspirante.persona,
                                      cohortemaestria=inscripcioncoohorte.cohortes,
                                      relacionados=None,
                                      nombre=tiporubroarancel.nombre + ' - ' + inscripcioncoohorte.cohortes.maestriaadmision.descripcion + ' - ' + inscripcioncoohorte.cohortes.descripcion,
                                      cuota=1,
                                      fecha=datetime.now().date(),
                                      fechavence=datetime.now().date() + timedelta(days=5),
                                      valor=inscripcioncoohorte.cohortes.valorexamen,
                                      iva_id=1,
                                      valoriva=0,
                                      valortotal=inscripcioncoohorte.cohortes.valorexamen,
                                      saldo=inscripcioncoohorte.cohortes.valorexamen,
                                      epunemi=True,
                                      admisionposgradotipo=1,
                                      idrubroepunemi=0,
                                      cancelado=False)
                        rubro.save(request)
                log(u'Envio estado aprobacion: %s' % (inscripcioncoohorte), request, "add")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al aprobar."})

        elif action == 'aprobarestadocomprobante':
            try:
                inscripcioncoohorte = EvidenciaPagoExamen.objects.get(
                    inscripcioncohorte_id=request.POST['idinscripcioncomprobanteestado'])
                # inscripcioncoohorte.fecha_aprobador = datetime.now()
                inscripcioncoohorte.estadorevision = request.POST['estadoaprobador']
                # inscripcioncoohorte.persona_aprobador = persona
                inscripcioncoohorte.save(request)
                log(u'Envio estado aprobacion: %s' % (inscripcioncoohorte), request, "add")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al aprobar."})

        elif action == 'aprobarnotificacionexamen':
            try:
                browser = request.POST['navegador']
                ops = request.POST['os']
                cookies = request.POST['cookies']
                screensize = request.POST['screensize']
                grupoexamen = GrupoExamenMsc.objects.get(pk=request.POST['idgrupoexamen'])
                grupoexamen.fecha_emailentrevista = datetime.now()
                grupoexamen.estado_emailentrevista = 2
                grupoexamen.persona_emailentrevista = persona
                grupoexamen.save(request)
                log(u'Envio email notificacion examen: %s' % (grupoexamen), request, "add")
                for integrantesentrevista in grupoexamen.integrantegrupoexamenmsc_set.filter(status=True):
                    #     send_html_mail("Aprobado Admision-UNEMI.", "emails/notificacionentrevista.html", {'sistema': u'Admision - UNEMI',  'fecha': datetime.now().date(), 'hora': datetime.now().time(), 'bs': browser, 'os': ops, 'cookies': cookies, 'screensize': screensize, 't': miinstitucion()}, integrantesentrevista.inscripcion.inscripcionaspirante.persona.emailpersonal(), [], cuenta=variable_valor('CUENTAS_CORREOS')[4])
                    send_html_mail("Examen Admision-UNEMI.", "emails/notificacionentrevista.html",
                                   {'sistema': u'Admision - UNEMI', 'fecha': datetime.now().date(),
                                    'hora': datetime.now().time(), 'bs': browser, 'os': ops, 'cookies': cookies,
                                    'screensize': screensize, 't': miinstitucion()},
                                   integrantesentrevista.inscripcion.inscripcionaspirante.persona.emailpersonal(), [],
                                   cuenta=variable_valor('CUENTAS_CORREOS')[16])
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al aprobar."})

        elif action == 'aprobarnotificacionexamenindividual':
            try:
                browser = request.POST['navegador']
                ops = request.POST['os']
                cookies = request.POST['cookies']
                screensize = request.POST['screensize']
                arregloemail = [23, 24, 25, 26, 27, 28]
                emailaleatorio = random.choice(arregloemail)
                inscripcionexamen = IntegranteGrupoExamenMsc.objects.get(pk=request.POST['inscritoexamen'])
                send_html_mail("Examen Admision-UNEMI.", "emails/notificacionentrevista.html",
                               {'sistema': u'Admision - UNEMI', 'fecha': datetime.now().date(),
                                'hora': datetime.now().time(), 'bs': browser, 'os': ops, 'cookies': cookies,
                                'screensize': screensize, 't': miinstitucion()},
                               inscripcionexamen.inscripcion.inscripcionaspirante.persona.emailpersonal(), [],
                               cuenta=variable_valor('CUENTAS_CORREOS')[emailaleatorio])
                time.sleep(8)
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al aprobar."})

        elif action == 'listadoentrevistado':
            try:
                grupoentrevista = GrupoEntrevistaMsc.objects.get(pk=request.POST['idgrupoentrevista'])
                grupoentrevista.estado_emailentrevista = 2
                grupoentrevista.persona_emailentrevista = persona
                grupoentrevista.save(request)
                listagrupoentrevista = grupoentrevista.integrantegrupoentrevitamsc_set.values('id',
                                                                                              'inscripcion__inscripcionaspirante__persona_id',
                                                                                              'inscripcion__inscripcionaspirante__persona__apellido1',
                                                                                              'inscripcion__inscripcionaspirante__persona__apellido2',
                                                                                              'inscripcion__inscripcionaspirante__persona__nombres').filter(
                    status=True).order_by('inscripcion__inscripcionaspirante__persona__apellido1',
                                          'inscripcion__inscripcionaspirante__persona__apellido2',
                                          'inscripcion__inscripcionaspirante__persona__nombres')
                return JsonResponse({"result": "ok", "cantidad": len(listagrupoentrevista),
                                     "listagrupoentrevista": list(listagrupoentrevista)})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al aprobar."})

        elif action == 'aprobarnotificacionentrevistaindividual':
            try:
                browser = request.POST['navegador']
                ops = request.POST['os']
                cookies = request.POST['cookies']
                screensize = request.POST['screensize']
                inscripcionentrevista = IntegranteGrupoEntrevitaMsc.objects.get(pk=request.POST['inscritoentrevista'])
                fechaentrevista = inscripcionentrevista.grupoentrevista.fecha
                arregloemail = [23, 24, 25, 26, 27, 28]
                emailaleatorio = random.choice(arregloemail)
                send_html_mail("Examen Admision-UNEMI.", "emails/notificacionentrevistagrupos.html",
                               {'sistema': u'Admision - UNEMI', 'fecha': datetime.now().date(),
                                'fechaentrevista': fechaentrevista, 'hora': datetime.now().time(), 'bs': browser,
                                'os': ops, 'cookies': cookies, 'screensize': screensize, 't': miinstitucion()},
                               inscripcionentrevista.inscripcion.inscripcionaspirante.persona.emailpersonal(), [],
                               cuenta=variable_valor('CUENTAS_CORREOS')[emailaleatorio])
                time.sleep(8)
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al aprobar."})

        elif action == 'aprobarnotificacionentrevista':
            try:
                browser = request.POST['navegador']
                ops = request.POST['os']
                cookies = request.POST['cookies']
                screensize = request.POST['screensize']
                grupoentrevista = GrupoEntrevistaMsc.objects.get(pk=request.POST['idgrupoentrevista'])
                grupoentrevista.fecha_emailentrevista = datetime.now()
                grupoentrevista.estado_emailentrevista = 2
                grupoentrevista.persona_emailentrevista = persona
                grupoentrevista.save(request)
                log(u'Envio email notificacion entrevista: %s' % (grupoentrevista), request, "add")
                for integrantesentrevista in grupoentrevista.integrantegrupoentrevitamsc_set.filter(status=True):
                    #     send_html_mail("Aprobado Admision-UNEMI.", "emails/notificacionentrevista.html", {'sistema': u'Admision - UNEMI',  'fecha': datetime.now().date(), 'hora': datetime.now().time(), 'bs': browser, 'os': ops, 'cookies': cookies, 'screensize': screensize, 't': miinstitucion()}, integrantesentrevista.inscripcion.inscripcionaspirante.persona.emailpersonal(), [], cuenta=variable_valor('CUENTAS_CORREOS')[4])
                    send_html_mail("Examen Admision-UNEMI.", "emails/notificacionentrevistagrupos.html",
                                   {'sistema': u'Admision - UNEMI', 'fecha': datetime.now().date(),
                                    'hora': datetime.now().time(), 'bs': browser, 'os': ops, 'cookies': cookies,
                                    'screensize': screensize, 't': miinstitucion()},
                                   integrantesentrevista.inscripcion.inscripcionaspirante.persona.emailpersonal(), [],
                                   cuenta=variable_valor('CUENTAS_CORREOS')[16])
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al aprobar."})

        elif action == 'inscribircarrera':
            try:
                f = InscripcionCarreraForm(request.POST)
                integrante = IntegranteGrupoEntrevitaMsc.objects.get(pk=int(encrypt(request.POST['id'])))
                if f.is_valid() and not Inscripcion.objects.filter(
                        persona=integrante.inscripcion.inscripcionaspirante.persona,
                        carrera=f.cleaned_data['carrera']).exists():
                    carrera = f.cleaned_data['carrera']
                    sesion = f.cleaned_data['sesion']
                    modalidad = f.cleaned_data['modalidad']
                    sede = f.cleaned_data['sede']
                    if Inscripcion.objects.filter(persona=integrante.inscripcion.inscripcionaspirante.persona,
                                                  carrera=carrera).exists():
                        return JsonResponse({"result": "bad", "mensaje": u"Ya se encuentra registrado en esa carrera."})
                    if not Inscripcion.objects.filter(persona=integrante.inscripcion.inscripcionaspirante.persona,
                                                      carrera=carrera).exists():
                        inscripcion = Inscripcion(persona=integrante.inscripcion.inscripcionaspirante.persona,
                                                  fecha=datetime.now().date(),
                                                  carrera=carrera,
                                                  modalidad=modalidad,
                                                  sesion=sesion,
                                                  sede=sede,
                                                  colegio='',
                                                  aplica_b2=True,
                                                  fechainicioprimernivel=datetime.now().date(),
                                                  fechainiciocarrera=datetime.now().date())
                        inscripcion.save(request)
                        integrante.inscripcion.inscripcionaspirante.persona.crear_perfil(inscripcion=inscripcion)
                        documentos = DocumentosDeInscripcion(inscripcion=inscripcion,
                                                             titulo=False,
                                                             acta=False,
                                                             cedula=False,
                                                             votacion=False,
                                                             actaconv=False,
                                                             partida_nac=False,
                                                             pre=False,
                                                             observaciones_pre='',
                                                             fotos=False)
                        documentos.save()
                        preguntasinscripcion = inscripcion.preguntas_inscripcion()
                        inscripciontesdrive = InscripcionTesDrive(inscripcion=inscripcion,
                                                                  licencia=False,
                                                                  record=False,
                                                                  certificado_tipo_sangre=False,
                                                                  prueba_psicosensometrica=False,
                                                                  certificado_estudios=False)
                        inscripciontesdrive.save()
                        # inscripcion.mi_malla()
                        inscripcion.malla_inscripcion()
                        inscripcion.actualizar_nivel()
                        if USA_TIPOS_INSCRIPCIONES:
                            inscripciontipoinscripcion = InscripcionTipoInscripcion(inscripcion=inscripcion,
                                                                                    tipoinscripcion_id=TIPO_INSCRIPCION_INICIAL)
                            inscripciontipoinscripcion.save()
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'rechazarnotificacion':
            try:
                browser = request.POST['navegador']
                ops = request.POST['os']
                cookies = request.POST['cookies']
                screensize = request.POST['screensize']
                inscripcioncoohorte = InscripcionCohorte.objects.get(pk=request.POST['idinscripcioncohorte'])
                inscripcioncoohorte.fecha_emailevidencia = datetime.now()
                inscripcioncoohorte.estado_emailevidencia = 3
                inscripcioncoohorte.persona_emailevidencia = persona
                inscripcioncoohorte.save(request)
                log(u'Envio email rechazado: %s' % (inscripcioncoohorte), request, "add")
                send_html_mail("Rechazado Admision-UNEMI.", "emails/registrorechazado.html",
                               {'sistema': u'Admision - UNEMI', 'fecha': datetime.now().date(),
                                'hora': datetime.now().time(), 'bs': browser, 'os': ops, 'cookies': cookies,
                                'screensize': screensize, 't': miinstitucion()},
                               inscripcioncoohorte.inscripcionaspirante.persona.emailpersonal(), [],
                               cuenta=variable_valor('CUENTAS_CORREOS')[16])
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al rechazar."})

        elif action == 'editarmaestriaadmision':
            try:
                f = AdmiPeriodoForm(request.POST)
                maestriaadmision = MaestriasAdmision.objects.get(pk=int(encrypt(request.POST['id'])))
                if f.is_valid():
                    maestriaadmision.descripcion = f.cleaned_data['descripcion']
                    if not maestriaadmision.en_uso():
                        maestriaadmision.carrera = f.cleaned_data['carrera']
                    maestriaadmision.save(request)
                    log(u'Editó maestría de admisión: %s' % maestriaadmision, request, "edit")
                    return JsonResponse({"result": False}, safe=False)
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'editarentrevistador':
            try:
                f = EntrevistadorCohorteForm(request.POST)
                grupoentrevista = GrupoEntrevistaMsc.objects.get(pk=int(encrypt(request.POST['id'])))
                if f.is_valid():
                    grupoentrevista.administrativo_id = f.cleaned_data['administrativo']
                    grupoentrevista.save(request)
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'editarrequisitomaestria':
            try:
                if 'archivo' in request.FILES:
                    newfile = request.FILES['archivo']
                    if newfile.size > 10485760:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Error, Tamaño de archivo Maximo permitido es de 10Mb"})
                f = AdmiRequisitosMaestriaForm(request.POST)
                requisito = RequisitosMaestria.objects.get(pk=int(encrypt(request.POST['id'])))
                if f.is_valid():
                    requisito.descripcion = f.cleaned_data['descripcion']
                    requisito.observacion = f.cleaned_data['observacion']
                    requisito.activo = f.cleaned_data['activo']
                    requisito.requerido = f.cleaned_data['requerido']
                    requisito.save(request)
                    if 'archivo' in request.FILES:
                        newfile = request.FILES['archivo']
                        newfile._name = generar_nombre("formato_", newfile._name)
                        requisito.archivo = newfile
                        requisito.save(request)
                    log(u'Editó requisito de maestría: %s' % requisito, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'editarpreguntamaestria':
            try:
                f = AdmiPreguntasMaestriaForm(request.POST)
                requisito = PreguntaMaestria.objects.get(pk=int(encrypt(request.POST['id'])))
                if f.is_valid():
                    requisito.descripcion = f.cleaned_data['descripcion']
                    requisito.activo = f.cleaned_data['activo']
                    requisito.save(request)
                    log(u'Editó pregunta de maestría: %s' % requisito, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'delprogramamaestria':
            try:
                maestria = CohorteMaestria.objects.get(pk=int(encrypt(request.POST['id'])))
                maestria.status = False
                maestria.save()
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'addinscritos':
            try:
                listainscritos = request.POST['inscritos']
                idgrupoentrevista = request.POST['idgrupoentrevista']
                for listado in listainscritos.split(':'):
                    integrantesentrevista = IntegranteGrupoEntrevitaMsc(grupoentrevista_id=idgrupoentrevista,
                                                                        inscripcion_id=listado[0])
                    integrantesentrevista.save(request)
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'addrespuestaspreguntas':
            try:
                listapreguntas = request.POST['listapreguntas']
                integrante = IntegranteGrupoEntrevitaMsc.objects.get(pk=request.POST['idintegrante'], status=True)
                integrante.estadoentrevista_id = request.POST['estadopregunta']
                integrante.save(request)
                if integrante.estadoentrevista.estado == 1:
                    inscrito = integrante.inscripcion
                    inscrito.aproboproceso = True
                    inscrito.save(request)
                else:
                    inscrito = integrante.inscripcion
                    inscrito.aproboproceso = False
                    inscrito.save(request)
                for listado in listapreguntas.split(':'):
                    cadena = listado.split(',')
                    if not RespuestaEntrevitaMsc.objects.filter(integrante=integrante, preguntacohorte_id=cadena[0],
                                                                status=True).exists():
                        respuesta = RespuestaEntrevitaMsc(integrante=integrante,
                                                          preguntacohorte_id=cadena[0],
                                                          respuesta=cadena[1])
                        respuesta.save(request)
                    else:
                        return JsonResponse({"result": "bad", "mensaje": u"Ya existen datos ingresados."})
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'editrespuestaspreguntas':
            try:
                listarespuestas = request.POST['listarespuesta']
                integrante = IntegranteGrupoEntrevitaMsc.objects.get(pk=request.POST['idintegrante'], status=True)
                integrante.estadoentrevista_id = request.POST['estadopregunta']
                integrante.save(request)
                if integrante.estadoentrevista.estado == 1:
                    inscrito = integrante.inscripcion
                    inscrito.aproboproceso = True
                    inscrito.save(request)
                else:
                    inscrito = integrante.inscripcion
                    inscrito.aproboproceso = False
                    inscrito.save(request)
                for respuestas in listarespuestas.split(':'):
                    cadena = respuestas.split(',')
                    respuesta = RespuestaEntrevitaMsc.objects.get(pk=cadena[0])
                    respuesta.respuesta = cadena[1]
                    respuesta.save(request)
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'addpreguntacohorte':
            try:
                listapreguntas = request.POST['listapreguntas']
                idcohorte = request.POST['idcohorte']
                for pregunta in listapreguntas.split(':'):
                    preguntamaestria = PreguntaMaestria(cohortes_id=idcohorte,
                                                        pregunta_id=int(pregunta))
                    preguntamaestria.save(request)
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'planificarhorarioentrevista':
            try:
                integrantes = IntegranteGrupoEntrevitaMsc.objects.get(pk=request.POST['idcodigo'])
                integrantes.fecha = request.POST['fecha']
                integrantes.horadesde = request.POST['id_horadesde']
                integrantes.horahasta = request.POST['id_horahasta']
                integrantes.lugar = request.POST['id_lugar']
                integrantes.save(request)
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'inscribirintegrantesentrevista':
            try:
                f = AdmiPagoMatriculaForm(request.POST)
                if f.is_valid():
                    inscritocohorte = InscripcionCohorte.objects.get(pk=encrypt(request.POST['id']))
                    inscritocohorte.fecharecibo = f.cleaned_data['fecha']
                    inscritocohorte.save(request)
                    pagos = Pago(inscripcioncohorte=inscritocohorte, tipo_id=3)
                    pagos.save(request)
                    detallepagos = CuotaPago(pago=pagos, fechapago=f.cleaned_data['fecha'],
                                             valor=f.cleaned_data['valor'])
                    detallepagos.save(request)
                    inscripcion = Inscripcion(persona=inscritocohorte.inscripcionaspirante.persona,
                                              fecha=inscritocohorte.fecharecibo,
                                              fechainicioprimernivel=inscritocohorte.fecharecibo,
                                              carrera=inscritocohorte.cohortes.maestriaadmision.carrera,
                                              coordinacion_id=18,
                                              modalidad=inscritocohorte.cohortes.modalidad,
                                              sesion_id=1,
                                              status=True,
                                              colegio='',
                                              activo=True,
                                              estado_gratuidad=1,
                                              identificador='',
                                              gratuidad=True,
                                              inscripcioncohorte=inscritocohorte,
                                              porcentaje_perdida_gratuidad=0,
                                              sede_id=1)
                    inscripcion.save()
                    if not inscritocohorte.inscripcionaspirante.persona.es_estudiante():
                        inscritocohorte.inscripcionaspirante.persona.crear_perfil(inscripcion=inscripcion)
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'delrequisitomaestria':
            try:
                maestria = RequisitosMaestria.objects.get(pk=int(encrypt(request.POST['id'])))
                maestria.status = False
                maestria.save()
                return JsonResponse({"result": "ok", "mensaje": u"Requisito eliminado correctamente"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'delpreguntacohorte':
            try:
                preguntacohorte = PreguntaMaestria.objects.get(pk=int(encrypt(request.POST['id'])))
                preguntacohorte.status = False
                preguntacohorte.save()
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'updatefecha':
            try:
                integrante = IntegranteGrupoEntrevitaMsc.objects.get(pk=request.POST['mid'])
                fechainicio = convertir_fecha(request.POST['fecha'])
                integrante.fecha = fechainicio
                integrante.save(request)
                return JsonResponse({'result': 'ok', 'fecha': integrante.fecha.strftime("%d-%m-%Y")})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad'})

        elif action == 'delpreguntamaestria':
            try:
                pregunta = PreguntaMaestria.objects.get(pk=int(encrypt(request.POST['id'])))
                pregunta.status = False
                pregunta.save()
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'addrequisito':
            try:
                form = RequisitoForm(request.POST)
                newfile = None
                if 'archivo' in request.FILES:
                    newfile = request.FILES['archivo']
                    if newfile:
                        if newfile.size > 10485760:
                            return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 10 Mb."})
                        else:
                            newfilesd = newfile._name
                            ext = newfilesd[newfilesd.rfind("."):]
                            if ext == '.pdf' or ext == '.PDF' or ext == '.doc' or ext == '.DOC' or ext == '.docx' or ext == '.DOCX':
                                newfile._name = generar_nombre("requisitos_", newfile._name)
                            else:
                                return JsonResponse(
                                    {"result": "bad", "mensaje": u"Error, Solo archivo con extención. pdf."})
                if form.is_valid():
                    requisito = Requisito(nombre=form.cleaned_data['nombre'],
                                          observacion=form.cleaned_data['observacion'],
                                          activo=form.cleaned_data['activo'],
                                          tipoarchivo=form.cleaned_data['tipoarchivo'],
                                          archivo=newfile)
                    requisito.save(request)
                    # clasificacion
                    for c in form.cleaned_data['clasificacion']:
                        clasereq = ClaseRequisito(requisito=requisito,
                                                  clasificacion=c)
                        clasereq.save(request)
                    log(u'Adiciono nuevo requisito: %s' % requisito, request, "add")
                    return JsonResponse({"result": False}, safe=False)
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        if action == 'eliminarequisito':
            try:
                idrequisito = request.POST['id']
                requisito = Requisito.objects.get(pk=idrequisito)
                requisito.status = False
                requisito.save()
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar requisito."})

        if action == 'deleteformatopreinscrito':
            try:
                idforca = request.POST['id']
                formato = FormatoCarreraIpec.objects.get(pk=idforca)
                formato.status = False
                formato.save()
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar requisito."})

        elif action == 'addpregunta':
            try:
                form = PreguntaForm(request.POST)
                if form.is_valid():
                    pregunta = PreguntasPrograma(tipopregunta=form.cleaned_data['tipopregunta'],
                                                 descripcion=form.cleaned_data['descripcion'])
                    pregunta.save(request)
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'addobservacionpreinscrito':
            try:
                form = EvidenciasMaestriasForm(request.POST)
                if form.is_valid():
                    evidencias = EvidenciasMaestrias.objects.get(pk=int(encrypt(request.POST['id'])))
                    evidencias.observaciones = form.cleaned_data['observaciones']
                    evidencias.save(request)
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'editpregunta':
            try:
                form = PreguntaForm(request.POST)
                if form.is_valid():
                    pregunta = PreguntasPrograma.objects.get(pk=int(encrypt(request.POST['id'])))
                    pregunta.descripcion = form.cleaned_data['descripcion']
                    pregunta.tipopregunta_id = request.POST['tipopregunta']
                    pregunta.save(request)
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'delpregunta':
            try:
                pregunta = PreguntasPrograma.objects.get(pk=int(encrypt(request.POST['id'])))
                pregunta.status = False
                pregunta.save()
                # if not requisito.requisitosmaestria_set.filter(status=True).exists():
                #     requisito.delete()
                # else:
                #     return JsonResponse({"result": "bad", "mensaje": u"Error Este registro esta sienddo usado."})
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'addtipopregunta':
            try:
                form = TipoPreguntaForm(request.POST)
                if form.is_valid():
                    tipopregunta = TipoPreguntasPrograma(descripcion=form.cleaned_data['descripcion'])
                    tipopregunta.save(request)
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'edittipopregunta':
            try:
                form = TipoPreguntaForm(request.POST)
                if form.is_valid():
                    tipopregunta = TipoPreguntasPrograma.objects.get(pk=int(encrypt(request.POST['id'])))
                    tipopregunta.descripcion = form.cleaned_data['descripcion']
                    tipopregunta.save(request)
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'deltipopregunta':
            try:
                tipopregunta = TipoPreguntasPrograma.objects.get(pk=int(encrypt(request.POST['id'])))
                tipopregunta.status = False
                tipopregunta.save()
                # if not requisito.requisitosmaestria_set.filter(status=True).exists():
                #     requisito.delete()
                # else:
                #     return JsonResponse({"result": "bad", "mensaje": u"Error Este registro esta sienddo usado."})
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'delinscritoentrevista':
            try:
                integrante = IntegranteGrupoEntrevitaMsc.objects.get(pk=int(encrypt(request.POST['id'])))
                integrante.status = False
                integrante.save()
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})


        elif action == 'editrequisito':
            try:
                form = RequisitoForm(request.POST)
                newfile = None
                if 'archivo' in request.FILES:
                    newfile = request.FILES['archivo']
                    if newfile:
                        if newfile.size > 10485760:
                            return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 10 Mb."})
                        else:
                            newfilesd = newfile._name
                            ext = newfilesd[newfilesd.rfind("."):]
                            if ext == '.pdf' or ext == '.PDF' or ext == '.doc' or ext == '.DOC' or ext == '.docx' or ext == '.DOCX':
                                newfile._name = generar_nombre("requisitos_", newfile._name)
                            else:
                                return JsonResponse(
                                    {"result": "bad", "mensaje": u"Error, Solo archivo con extención. pdf."})
                if form.is_valid():
                    requisito = Requisito.objects.get(pk=int(encrypt(request.POST['id'])))
                    requisito.nombre = form.cleaned_data['nombre']
                    requisito.observacion = form.cleaned_data['observacion']
                    requisito.tipoarchivo = form.cleaned_data['tipoarchivo']
                    requisito.activo = form.cleaned_data['activo']
                    if newfile:
                        requisito.archivo = newfile
                    requisito.save(request)

                    # clasificacion, validar de no quitar clasificacion de requisitos en uso
                    delclases = ClaseRequisito.objects.filter(requisito=requisito, status=True)
                    mensaje = str(requisito) + ': '
                    for d in delclases:
                        if not d.esta_uso():
                            d.status = False
                            d.save()
                        else:
                            if not d.clasificacion in form.cleaned_data['clasificacion']:
                                mensaje = mensaje + str(d.clasificacion) + ', '
                    for c in form.cleaned_data['clasificacion']:
                        if not ClaseRequisito.objects.filter(requisito=requisito, clasificacion=c,
                                                             status=True).exists():
                            clase = ClaseRequisito(requisito=requisito, clasificacion=c)
                            clase.save(request)
                    if mensaje != str(requisito) + ': ':
                        messages.warning(request,
                                         'Clasificación de ' + mensaje + ' estan en uso. No se pueden desclasificar.')

                    log(u'Editó el requisito: %s' % requisito, request, "edit")
                    return JsonResponse({"result": False}, safe=False)
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'delrequisito':
            try:
                requisito = Requisito.objects.get(pk=int(encrypt(request.POST['id'])))
                if not requisito.requisitosmaestria_set.filter(status=True).exists():
                    requisito.status = False
                    requisito.save()
                else:
                    return JsonResponse({"result": "bad", "mensaje": u"Error Este registro esta siendo usado."})
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})


        elif action == 'actualizar_requisito':
            try:
                estado = False
                requisito = Requisito.objects.get(status=True, pk=int(encrypt(request.POST['idr'])))
                if requisito.requisitosmaestria_set.filter(status=True, requisito=requisito,
                                                           cohorte_id=int(encrypt(request.POST['idc']))).exists():
                    rc = requisito.requisitosmaestria_set.get(status=True, requisito=requisito,
                                                              cohorte_id=int(encrypt(request.POST['idc'])))
                    rc.status = False
                    rc.save()
                else:
                    rc = RequisitosMaestria(requisito=requisito, cohorte_id=int(encrypt(request.POST['idc'])))
                    rc.save(request)
                    estado = True
                return JsonResponse({"result": "ok", "id": requisito.id, "estado": estado})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})


        elif action == 'listarequisito':
            try:
                data['title'] = u'Listado de Requisitos'
                data['cohorte'] = CohorteMaestria.objects.get(status=True, pk=int(encrypt(request.POST['id'])))
                data['requisitos'] = Requisito.objects.filter(activo=True, status=True)
                template = get_template("adm_admision/addrequisitomaestria.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'data': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})


        elif action == 'enviarcorreo':
            try:
                preinscrito = PreInscripcion.objects.get(pk=int(encrypt(request.POST['id'])))
                preinscrito.enviocorreo = True
                preinscrito.save(request)
                formatocorreo = FormatoCarreraIpec.objects.filter(carrera=preinscrito.carrera, status=True)[0]
                lista = []
                if preinscrito.persona.emailinst:
                    lista.append(preinscrito.persona.emailinst)
                if preinscrito.persona.email:
                    lista.append(preinscrito.persona.email)
                if formatocorreo.correomaestria:
                    lista.append(formatocorreo.correomaestria)
                lista.append(conectar_cuenta(CUENTAS_CORREOS[18][1]))

                asunto = u"Requisitos para admisión de " + preinscrito.carrera.nombre
                # asunto = u"Requisitos para admisión"
                send_html_mail(asunto, "emails/notificacion_requisito_ipec.html",
                               {'sistema': 'Posgrado UNEMI', 'preinscrito': preinscrito,
                                'formato': formatocorreo.banner},
                               lista, [], [formatocorreo.archivo],
                               cuenta=CUENTAS_CORREOS[18][1])
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": "%s" % ex})


        elif action == 'enviarcorreoaceptar':
            try:
                preinscrito = PreInscripcion.objects.get(pk=int(encrypt(request.POST['id'])))
                formato = FormatoCarreraIpec.objects.filter(carrera=preinscrito.carrera, status=True)[0]
                preinscrito.aceptarpreinscripcion = True
                preinscrito.save(request)
                lista = []
                if preinscrito.persona.emailinst:
                    lista.append(preinscrito.persona.emailinst)
                if preinscrito.persona.email:
                    lista.append(preinscrito.persona.email)

                lista.append(conectar_cuenta(CUENTAS_CORREOS[18][1]))

                asunto = u"Aceptación del proceso de admisión a la " + preinscrito.carrera.nombre
                send_html_mail(asunto, "emails/aceptar_preinscripcion_ipec.html",
                               {'sistema': 'Posgrado', 'preinscrito': preinscrito, 'formato': formato.banner},
                               lista, [], [],
                               cuenta=CUENTAS_CORREOS[18][1])
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": "%s" % ex})


        elif action == 'addgrupoexamen':
            try:
                form = GrupoExamenForm(request.POST)
                if form.is_valid():
                    cohorte = CohorteMaestria.objects.values("id").get(status=True, pk=int(encrypt(request.POST['id'])))
                    grupo = GrupoExamenMsc(cohorte_id=cohorte['id'],
                                           profesor=form.cleaned_data['profesor'],
                                           # profesor_id=int(form.cleaned_data['profesor']),
                                           lugar=form.cleaned_data['lugar'],
                                           fecha=form.cleaned_data['fecha'],
                                           hora=form.cleaned_data['hora'],
                                           visible=form.cleaned_data['visible'],
                                           urlzoom=form.cleaned_data['urlzoom'],
                                           observacion=form.cleaned_data['observacion'])
                    grupo.save(request)
                    log(u'Adiciono nuevo Grupo: %s' % grupo, request, "addgrupoexamen")
                    if 'lista_items1' in request.POST:
                        for idpar in json.loads(request.POST['lista_items1']):
                            if not IntegranteGrupoExamenMsc.objects.values("id").filter(status=True,
                                                                                        inscripcion_id=idpar,
                                                                                        grupoexamen=grupo).exists():
                                participante = IntegranteGrupoExamenMsc(inscripcion_id=idpar, grupoexamen=grupo, )
                                participante.save(request)
                return JsonResponse({"result": False}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'editgrupoexamen':
            try:
                form = GrupoExamenForm(request.POST)
                if form.is_valid():
                    grupo = GrupoExamenMsc.objects.get(status=True, pk=int(encrypt(request.POST['id'])),
                                                       cohorte_id=int(encrypt(request.POST['idcohorte'])))
                    grupo.profesor = form.cleaned_data['profesor']
                    grupo.lugar = form.cleaned_data['lugar']
                    grupo.fecha = form.cleaned_data['fecha']
                    grupo.hora = form.cleaned_data['hora']
                    grupo.urlzoom = form.cleaned_data['urlzoom']
                    grupo.visible = form.cleaned_data['visible']
                    grupo.observacion = form.cleaned_data['observacion']
                    grupo.save(request)
                return JsonResponse({"result": False}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'delintegrante':
            try:
                if 'id' in request.POST:
                    integrante = IntegranteGrupoExamenMsc.objects.get(status=True, pk=int(encrypt(request.POST['id'])))
                    if integrante.puede_eliminar_integrante():
                        integrante.status = False
                        integrante.save()
                        return JsonResponse({"result": "ok"})
                    else:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"No puede eliminar el integrante ya calificó el profesor."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'delgrupoexamen':
            try:
                if 'id' in request.POST:
                    integrante = GrupoExamenMsc.objects.get(status=True, pk=int(encrypt(request.POST['id'])))
                    if integrante.puede_eliminar_grupo():
                        integrante.status = False
                        integrante.save()
                        return JsonResponse({"result": "ok"})
                    else:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"No puede eliminar el grupo tiene integrantes activos."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'creargrupomoodle':
            try:
                if 'id' in request.POST:
                    grupo = GrupoExamenMsc.objects.get(status=True, pk=int(encrypt(request.POST['id'])))
                    grupo.crear_grupo_moodle()
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'addgrupoentrevista':
            try:
                form = GrupoEntrevistaForm(request.POST)
                if form.is_valid():
                    cohorte = CohorteMaestria.objects.values("id").get(status=True, pk=int(encrypt(request.POST['id'])))
                    grupo = GrupoEntrevistaMsc(cohortes_id=cohorte['id'],
                                               administrativo_id=int(form.cleaned_data['administrativo']),
                                               lugar=form.cleaned_data['lugar'],
                                               fecha=form.cleaned_data['fecha'],
                                               visible=form.cleaned_data['visible'],
                                               urlzoom=form.cleaned_data['urlzoom'],
                                               horainicio=form.cleaned_data['horainicio'],
                                               observacion=form.cleaned_data['observacion'])
                    grupo.save(request)
                    if 'lista_items1' in request.POST:
                        for idpar in json.loads(request.POST['lista_items1']):
                            if not IntegranteGrupoEntrevitaMsc.objects.values("id").filter(status=True,
                                                                                           inscripcion_id=idpar,
                                                                                           grupoentrevista=grupo).exists():
                                participante = IntegranteGrupoEntrevitaMsc(inscripcion_id=idpar,
                                                                           grupoentrevista=grupo,
                                                                           lugar=form.cleaned_data['lugar'],
                                                                           fecha=form.cleaned_data['fecha'])
                                participante.save(request)
                    return JsonResponse({"result": "ok"})
                else:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al guadar los datos."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'editgrupoentrevista':
            # from datetime import datetime, date, time, timedelta
            try:
                form = GrupoEntrevistaForm(request.POST)
                if form.is_valid():
                    grupo = GrupoEntrevistaMsc.objects.get(status=True, pk=int(encrypt(request.POST['id'])),
                                                           cohortes_id=int(encrypt(request.POST['idc'])))
                    grupo.administrativo_id = int(form.cleaned_data['administrativo'])
                    grupo.lugar = form.cleaned_data['lugar']
                    grupo.fecha = form.cleaned_data['fecha']
                    grupo.visible = form.cleaned_data['visible']
                    grupo.horainicio = form.cleaned_data['horainicio']
                    grupo.observacion = form.cleaned_data['observacion']
                    grupo.urlzoom = form.cleaned_data['urlzoom']
                    grupo.save(request)
                    varrango = grupo.cohortes.minutosrango
                    incrementar = 1
                    for listado in grupo.integrantegrupoentrevitamsc_set.filter(status=True).order_by(
                            'inscripcion__inscripcionaspirante__persona__apellido1',
                            'inscripcion__inscripcionaspirante__persona__apellido2'):
                        if incrementar == 1:
                            listado.horadesde = str(
                                timedelta(hours=grupo.horainicio.hour, minutes=grupo.horainicio.minute))
                            listado.save(request)
                        else:
                            varhorainicio = str(
                                timedelta(hours=grupo.horainicio.hour, minutes=grupo.horainicio.minute) + timedelta(
                                    hours=0, minutes=varrango))
                            varrango = varrango + grupo.cohortes.minutosrango
                            listado.horadesde = varhorainicio
                            listado.save(request)
                        incrementar = incrementar + 1
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'cargamasiva':
            try:
                inscripcioncohorte = InscripcionCohorte.objects.get(pk=request.POST['id'])
                listadorequisitosmaestria = RequisitosMaestria.objects.filter(cohorte=inscripcioncohorte.cohortes,
                                                                              status=True)
                for requi in listadorequisitosmaestria:
                    nombrefile = 'requisito' + str(requi.id)
                    if nombrefile in request.FILES:
                        if not EvidenciaRequisitosAspirante.objects.filter(requisitos=requi,
                                                                           inscripcioncohorte=inscripcioncohorte,
                                                                           status=True).exists():
                            requisitomaestria = EvidenciaRequisitosAspirante(requisitos=requi,
                                                                             inscripcioncohorte=inscripcioncohorte)
                            requisitomaestria.save(request)
                            newfile = request.FILES[nombrefile]
                            newfile._name = generar_nombre("requisitopgrado_" + str(requi.id) + "_", newfile._name)
                            requisitomaestria.archivo = newfile
                            requisitomaestria.save(request)
                            log(u'Adicionó requisito de maestria aspirante: %s' % requisitomaestria.requisitos, request,
                                "add")
                            detalle = DetalleEvidenciaRequisitosAspirante(evidencia=requisitomaestria,
                                                                          estadorevision=1,
                                                                          fecha=datetime.now().date(),
                                                                          observacion='')
                            detalle.save(request)
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'delformatocarrera':
            try:
                formato = FormatoCarreraIpec.objects.get(pk=int(encrypt(request.POST['id'])))
                formato.status = False
                formato.save(request)
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'editpreinscrito':
            # from datetime import datetime, date, time, timedelta
            try:
                form = PreInscripcionForm(request.POST)
                if form.is_valid():
                    preinscripcion = PreInscripcion.objects.get(status=True, pk=int(encrypt(request.POST['id'])))
                    preinscripcion.carrera = form.cleaned_data['carrera']
                    preinscripcion.save(request)
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'addfirma':

            try:
                from sga.forms import FirmaPersonaForm
                if not Administrativo.objects.filter(persona_id=request.POST['persona'], status=True,
                                                     persona__status=True).exists():
                    return JsonResponse(
                        {"result": "bad", "mensaje": u"Error, esta persona no se encuentra en administrativos."})

                if not FirmaPersona.objects.filter(persona_id=request.POST['persona'], tipofirma=2,
                                                   status=True).exists() and not 'firma' in request.FILES:
                    return JsonResponse(
                        {"result": "bad", "mensaje": u"Debe ingresar una imagen de su firma en blanco y negro."})

                adm = Administrativo.objects.get(persona_id=request.POST['persona'])
                configurarfirma = ConfigurarFirmaAdmisionPosgrado(
                    administrativo_id=adm.id,
                    cargo=request.POST['cargo'].upper()
                )
                configurarfirma.save(request)
                if 'firma' in request.FILES:
                    arch = request.FILES['firma']
                    extension = arch._name.split('.')
                    tam = len(extension)
                    exte = extension[tam - 1]
                    if arch.size > 6291456:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Error, el tamaño del archivo es mayor a 6 Mb."})
                    if not exte.lower() == 'png':
                        return JsonResponse({"result": "bad", "mensaje": u"Solo se permiten archivos .png"})

                    # form = FirmaPersonaForm(request.POST, request.FILES)
                    # if form.is_valid():
                    firmapersona = None
                    if not FirmaPersona.objects.filter(persona_id=int(request.POST['persona']), tipofirma=2,
                                                       status=True).exists():
                        firmapersona = FirmaPersona(persona_id=int(request.POST['persona']), tipofirma=2)
                        firmapersona.save(request)
                    else:
                        firmapersona = FirmaPersona.objects.filter(persona_id=int(request.POST['persona']), tipofirma=2,
                                                                   status=True).first()

                    ruta = os.path.join(SITE_STORAGE, 'media', 'reportes', 'encabezados_pies', 'firmas', '')
                    rutapdf = ruta + u"%s_%s.png" % (firmapersona.persona.cedula, firmapersona.tipofirma)
                    if os.path.isfile(rutapdf):
                        os.remove(rutapdf)

                    archivofirma = request.FILES['firma']
                    archivofirma._name = u"%s_%s.png" % (firmapersona.persona.cedula, firmapersona.tipofirma)
                    firmapersona.firma = archivofirma
                    firmapersona.save(request)

                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'editfirma':
            try:
                from sga.forms import FirmaPersonaForm
                cf = ConfigurarFirmaAdmisionPosgrado.objects.get(pk=int(encrypt(request.POST['id'])))
                cf.administrativo = Administrativo.objects.get(persona_id=request.POST['persona'])
                cf.cargo = request.POST['cargo'].strip().upper()
                cf.save(request)
                form = FirmaPersonaForm(request.POST, request.FILES)
                if 'firma' in request.FILES:
                    arch = request.FILES['firma']
                    extension = arch._name.split('.')
                    tam = len(extension)
                    exte = extension[tam - 1]
                    if arch.size > 6291456:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Error, el tamaño del archivo es mayor a 6 Mb."})
                    if not exte.lower() == 'png':
                        return JsonResponse({"result": "bad", "mensaje": u"Solo se permiten archivos .png"})
                firmapersona = FirmaPersona.objects.get(pk=int(encrypt(request.POST['idf'])))
                if 'firma' in request.FILES:
                    ruta = os.path.join(SITE_STORAGE, 'media', 'reportes', 'encabezados_pies', 'firmas', '')
                    rutapdf = ruta + u"%s_%s.png" % (firmapersona.persona.cedula, firmapersona.tipofirma)
                    if os.path.isfile(rutapdf):
                        os.remove(rutapdf)

                    archivofirma = request.FILES['firma']
                    archivofirma._name = u"%s_%s.png" % (firmapersona.persona.cedula, firmapersona.tipofirma)
                    firmapersona.firma = archivofirma
                    firmapersona.save(request)

                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'editformatocarrera':
            try:
                form = FormatoCarreraForm(request.POST)
                newfile = None
                newfile1 = None
                if 'banner' in request.FILES:
                    newfile1 = request.FILES['banner']
                    if newfile1:
                        if newfile1.size > 2097152:
                            return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 20 Mb."})
                        else:
                            newfilesd1 = newfile1._name
                            ext1 = newfilesd1[newfilesd1.rfind("."):]
                            if ext1 == '.jpg' or ext1 == '.png':
                                newfile1._name = generar_nombre("banner_", newfile1._name)
                            else:
                                return JsonResponse(
                                    {"result": "bad", "mensaje": u"Error, Solo archivo con extención. jpg."})
                if 'archivo' in request.FILES:
                    newfile = request.FILES['archivo']
                    if newfile:
                        if newfile.size > 26214400:
                            return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 20 Mb."})
                        else:
                            newfilesd = newfile._name
                            ext = newfilesd[newfilesd.rfind("."):]
                            if ext == '.pdf' or ext == '.PDF' or ext == '.doc' or ext == '.DOC' or ext == '.docx' or ext == '.DOCX':
                                newfile._name = generar_nombre("formatopreinscrito2_", newfile._name)
                            else:
                                return JsonResponse(
                                    {"result": "bad", "mensaje": u"Error, Solo archivo con extención. pdf."})
                if form.is_valid():

                    formato = FormatoCarreraIpec.objects.get(pk=int(encrypt(request.POST['id'])))
                    formato.correomaestria = form.cleaned_data['correomaestria']
                    formato.carrera = form.cleaned_data['carrera']

                    if newfile:
                        formato.archivo = newfile
                    if newfile1:
                        formato.banner = newfile1
                    formato.save(request)
                    log(u'Editó el requisito: %s' % formato, request, "edit")
                    return JsonResponse({"result": False}, safe=False)
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'delpreinscrito':
            try:
                preinscripcion = PreInscripcion.objects.get(pk=int(encrypt(request.POST['id'])))
                preinscripcion.status = False
                preinscripcion.save(request)
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})


        elif action == 'addformatopreinscrito':
            try:
                form = FormatoCarreraForm(request.POST)
                newfile = None
                if 'archivo' in request.FILES:
                    newfile = request.FILES['archivo']
                    if newfile:
                        if newfile.size > 50485760:
                            return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 20 Mb."})
                        else:
                            newfilesd = newfile._name
                            ext = newfilesd[newfilesd.rfind("."):]
                            if ext == '.pdf' or ext == '.PDF' or ext == '.doc' or ext == '.DOC' or ext == '.docx' or ext == '.DOCX':
                                newfile._name = generar_nombre("formatopreinscrito_", newfile._name)
                            else:
                                return JsonResponse(
                                    {"result": "bad", "mensaje": u"Error, Solo archivo con extención. pdf."})
                if form.is_valid():
                    if FormatoCarreraIpec.objects.filter(carrera=form.cleaned_data['carrera'], status=True).exists():
                        return JsonResponse({"result": "bad", "mensaje": u"Carrera ya esta registrada"})
                    formato = FormatoCarreraIpec(carrera=form.cleaned_data['carrera'],
                                                 correomaestria=form.cleaned_data['correomaestria'],
                                                 archivo=newfile)
                    formato.save(request)
                    log(u'Adiciono nuevo requisito: %s' % formato, request, "add")
                    return JsonResponse({"result": False}, safe=False)
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'delintegranteentrevista':
            try:
                if 'id' in request.POST:
                    integrante = IntegranteGrupoEntrevitaMsc.objects.get(status=True,
                                                                         pk=int(encrypt(request.POST['id'])))
                    if not integrante.tiene_entrevista():
                        integrante.status = False
                        integrante.save()
                        return JsonResponse({"result": "ok"})
                    else:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"No puede eliminar el integrante ya ha sido entrevistado."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'delgrupoentrevista':
            try:
                if 'id' in request.POST:
                    grupo = GrupoEntrevistaMsc.objects.get(status=True, pk=int(encrypt(request.POST['id'])))
                    if grupo.puede_eliminar_grupo_entrevista():
                        grupo.status = False
                        grupo.save()
                        return JsonResponse({"result": "ok"})
                    else:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"No puede eliminar el grupo tiene integrantes activos."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'consultarevidencia':
            try:
                if 'id' in request.POST:
                    evidencia = EvidenciaRequisitosAspirante.objects.get(status=True,
                                                                         pk=int(encrypt(request.POST['id'])))
                    detalle = evidencia.ultima_evidencia_aprobador()
                    return JsonResponse(
                        {"result": "ok", "evidencia": "Evidencia: " + str(evidencia.requisitos.requisito.nombre),
                         "obs": detalle.observacion if detalle.observacion else ""})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'verificarevidencia':
            try:
                if 'id' in request.POST:
                    evidencia = EvidenciaRequisitosAspirante.objects.get(status=True, pk=int(request.POST['id']))
                    detalle = evidencia.ultima_evidencia()
                    return JsonResponse({"result": "ok", "id": evidencia.id, "estado": detalle.estadorevision,
                                         "estadon": str(detalle.get_estadorevision_display()),
                                         "fechaaprobacion": detalle.fecha.strftime("%d/%m/%Y"),
                                         "comentario": detalle.observacion, "aprobador": str(detalle.persona)})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'misevidencia':
            try:
                if 'idr' in request.POST and 'idc' in request.POST and 'idi' in request.POST:
                    evidencia = EvidenciaRequisitosAspirante.objects.get(status=True,
                                                                         inscripcioncohorte__cohortes_id=int(
                                                                             request.POST['idc']),
                                                                         inscripcioncohorte_id=int(request.POST['idi']),
                                                                         requisitos_id=int(request.POST['idr']))
                    return JsonResponse({"result": "ok", "id": evidencia.id})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'aprobarrequisito':
            try:
                if 'id' in request.POST:
                    evidencia = EvidenciaRequisitosAspirante.objects.get(status=True,
                                                                         pk=int(encrypt(request.POST['id'])))
                    detalle = DetalleEvidenciaRequisitosAspirante(evidencia=evidencia,
                                                                  estadorevision=2,
                                                                  persona=persona,
                                                                  fecha=datetime.now().date(),
                                                                  observacion='')
                    detalle.save(request)
                    return JsonResponse({"result": "ok", "id": evidencia.id, "estado": detalle.estadorevision,
                                         "estadon": str(detalle.get_estadorevision_display()),
                                         "fechaaprobacion": detalle.fecha, "comentario": detalle.observacion,
                                         "aprobador": str(detalle.persona)})
                else:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al guadar los datos."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'rechazarrequisito':
            try:
                if 'id' in request.POST:
                    evidencia = EvidenciaRequisitosAspirante.objects.get(status=True,
                                                                         pk=int(encrypt(request.POST['id'])))
                    detalle = DetalleEvidenciaRequisitosAspirante(evidencia=evidencia,
                                                                  estadorevision=3,
                                                                  persona=persona,
                                                                  fecha=datetime.now().date(),
                                                                  observacion=request.POST['obs'])
                    detalle.save(request)
                    return JsonResponse({"result": "ok", "id": evidencia.id, "estado": detalle.estadorevision,
                                         "estadon": str(detalle.get_estadorevision_display()),
                                         "fechaaprobacion": detalle.fecha, "comentario": detalle.observacion,
                                         "aprobador": str(detalle.persona)})
                else:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al guadar los datos."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'generargrupoexamen':
            try:
                if 'id' in request.POST:
                    idintegrantes = []
                    cohorte = CohorteMaestria.objects.get(status=True, pk=int(encrypt(request.POST['id'])))
                    integrantes = cohorte.inscripcioncohorte_set.filter(status=True, activo=True, aproboproceso=False,
                                                                        integrantegrupoexamenmsc__isnull=True)
                    if integrantes:
                        for integrante in integrantes:
                            if integrante.total_evidenciasaprobadas() == cohorte.total_evidencia_cohorte():
                                idintegrantes.append([integrante.id])
                        if idintegrantes.__len__() > 0:
                            if cohorte.grupoexamenmsc_set.filter(status=True).exists():
                                grupo = cohorte.grupoexamenmsc_set.filter(status=True).order_by('-id')[0]
                                if grupo.integrantegrupoexamenmsc_set.values('id').filter(
                                        status=True).count() > cohorte.cantidadgruposexamen:
                                    grupo = GrupoExamenMsc(cohorte=cohorte, visible=False)
                                    grupo.save(request)
                                else:
                                    grupo = GrupoExamenMsc(cohorte=cohorte, visible=False)
                                    grupo.save(request)
                            else:
                                grupo = GrupoExamenMsc(cohorte=cohorte, visible=False)
                                grupo.save(request)
                            for idintegrante in idintegrantes:
                                integrante = IntegranteGrupoExamenMsc(inscripcion_id=idintegrante[0], grupoexamen=grupo)
                                integrante.save(request)
                            return JsonResponse({"result": "ok"})
                        else:
                            return JsonResponse({"result": "bad",
                                                 "mensaje": u"Los integrantes no cumplen con los requisistos para generar grupos"})
                    else:
                        return JsonResponse({"result": "bad", "mensaje": u"No existen integrantes en la corte" + str(
                            cohorte.descripcion)})
                else:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al guadar los datos."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'generaraprobadosgrupoexamen':
            try:
                if 'id' in request.POST:
                    cohorte = CohorteMaestria.objects.get(status=True, pk=int(encrypt(request.POST['id'])))
                    integrantes = cohorte.inscripcioncohorte_set.filter(status=True, activo=True,
                                                                        estado_emailevidencia=2, aproboproceso=False,
                                                                        integrantegrupoexamenmsc__isnull=True).order_by(
                        'fecha_emailevidencia')
                    f = integrantes.values('id').count()
                    if integrantes:
                        for integrante in integrantes:
                            if cohorte.grupoexamenmsc_set.filter(status=True).exists():
                                grupo = cohorte.grupoexamenmsc_set.filter(status=True).order_by('-id')[0]
                                if grupo.integrantegrupoexamenmsc_set.values('id').filter(
                                        status=True).count() >= cohorte.cantidadgruposexamen:
                                    grupo = GrupoExamenMsc(cohorte=cohorte, visible=False)
                                    grupo.save(request)
                                    integranteexamen = IntegranteGrupoExamenMsc(inscripcion=integrante,
                                                                                grupoexamen=grupo)
                                    integranteexamen.save(request)
                                else:
                                    integranteexamen = IntegranteGrupoExamenMsc(inscripcion=integrante,
                                                                                grupoexamen=grupo)
                                    integranteexamen.save(request)
                            else:
                                grupo = GrupoExamenMsc(cohorte=cohorte, visible=False)
                                grupo.save(request)
                                integranteexamen = IntegranteGrupoExamenMsc(inscripcion=integrante, grupoexamen=grupo)
                                integranteexamen.save(request)
                        return JsonResponse({"result": "ok"})
                        # if integrante.total_evidenciasaprobadas() == cohorte.total_evidencia_cohorte():
                        #     idintegrantes.append([integrante.id])
                        # if idintegrantes.__len__()>0:
                        #     if cohorte.grupoexamenmsc_set.filter(status=True).exists():
                        #         grupo = cohorte.grupoexamenmsc_set.filter(status=True).order_by('-id')[0]
                        #         if grupo.integrantegrupoexamenmsc_set.filter(status=True).count()>cohorte.cantidadgruposexamen:
                        #             grupo = GrupoExamenMsc(cohorte=cohorte, visible=False)
                        #             grupo.save(request)
                        #         else:
                        #             grupo = GrupoExamenMsc(cohorte=cohorte, visible=False)
                        #             grupo.save(request)
                        #     else:
                        #         grupo = GrupoExamenMsc(cohorte=cohorte, visible=False)
                        #         grupo.save(request)
                        #     for idintegrante in idintegrantes:
                        #         integrante = IntegranteGrupoExamenMsc(inscripcion_id=idintegrante[0], grupoexamen=grupo)
                        #         integrante.save(request)
                        #     return JsonResponse({"result": "ok"})
                        # else:
                        #     return JsonResponse({"result": "bad", "mensaje": u"Los integrantes no cumplen con los requisistos para generar grupos"})
                    else:
                        return JsonResponse({"result": "bad", "mensaje": u"No existen integrantes en la corte" + str(
                            cohorte.descripcion)})
                else:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al guadar los datos."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'generargrupoentrevista':
            try:
                if 'id' in request.POST:
                    cohorte = CohorteMaestria.objects.get(status=True, pk=int(encrypt(request.POST['id'])))
                    integrantes = cohorte.inscripcioncohorte_set.filter(status=True, activo=True,
                                                                        estado_emailevidencia=2, aproboproceso=False,
                                                                        integrantegrupoexamenmsc__estado=2,
                                                                        integrantegrupoentrevitamsc__isnull=True).order_by(
                        'inscripcionaspirante__persona__apellido1', 'inscripcionaspirante__persona__apellido2')
                    if integrantes:
                        for integrante in integrantes:
                            if cohorte.grupoentrevistamsc_set.filter(status=True).exists():
                                grupo = cohorte.grupoentrevistamsc_set.filter(status=True).order_by('-id')[0]
                                if grupo.integrantegrupoentrevitamsc_set.filter(
                                        status=True).count() >= cohorte.cantidadgruposentrevista:
                                    grupo = GrupoEntrevistaMsc(cohortes=cohorte, visible=False)
                                    grupo.save(request)
                                    integranteexamen = IntegranteGrupoEntrevitaMsc(inscripcion=integrante,
                                                                                   grupoentrevista=grupo)
                                    integranteexamen.save(request)
                                else:
                                    integranteexamen = IntegranteGrupoEntrevitaMsc(inscripcion=integrante,
                                                                                   grupoentrevista=grupo)
                                    integranteexamen.save(request)
                            else:
                                grupo = GrupoEntrevistaMsc(cohortes=cohorte, visible=False)
                                grupo.save(request)
                                integranteexamen = IntegranteGrupoEntrevitaMsc(inscripcion=integrante,
                                                                               grupoentrevista=grupo)
                                integranteexamen.save(request)
                        return JsonResponse({"result": "ok"})
                    else:
                        return JsonResponse({"result": "bad",
                                             "mensaje": u"No existen integrantes que aprobaron examen en la corte" + str(
                                                 cohorte.descripcion)})
                else:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al guadar los datos."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        if action == 'addhorainicio':
            try:
                if 'id' in request.POST:
                    integrante = IntegranteGrupoEntrevitaMsc.objects.get(status=True,
                                                                         pk=int(encrypt(request.POST['id'])),
                                                                         grupoentrevista_id=int(
                                                                             encrypt(request.POST['idg'])))
                    if not integrante.tiene_entrevista():
                        integrante.horadesde = request.POST['hora']
                        integrante.save(request)
                        return JsonResponse({"result": "ok", "id": integrante.id})
                    else:
                        return JsonResponse({"result": "bad",
                                             "mensaje": u"No puede modificar la hora inicio la entrevista ya ha sido tomada."})
                else:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al guadar los datos."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        if action == 'addhorafin':
            try:
                if 'id' in request.POST:
                    integrante = IntegranteGrupoEntrevitaMsc.objects.get(status=True,
                                                                         pk=int(encrypt(request.POST['id'])),
                                                                         grupoentrevista_id=int(
                                                                             encrypt(request.POST['idg'])))
                    if not integrante.tiene_entrevista():
                        if convertir_hora(request.POST['hora']) > integrante.horadesde:
                            integrante.horahasta = request.POST['hora']
                            integrante.save(request)
                            return JsonResponse({"result": "ok"})
                        else:
                            return JsonResponse({"result": "bad", "id": integrante.id,
                                                 "mensaje": u"Error hora debe ser mayor a la hora inicio."})
                    else:
                        return JsonResponse({"result": "bad",
                                             "mensaje": u"No puede modificar la hora fin la entrevista ya ha sido tomada."})
                else:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al guadar los datos."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        if action == 'addfecha':
            try:
                if 'id' in request.POST:
                    integrante = IntegranteGrupoEntrevitaMsc.objects.get(status=True,
                                                                         pk=int(encrypt(request.POST['id'])),
                                                                         grupoentrevista_id=int(
                                                                             encrypt(request.POST['idg'])))
                    if not integrante.tiene_entrevista():
                        integrante.fecha = convertir_fecha(request.POST['fecha'])
                        integrante.save(request)
                        return JsonResponse({"result": "ok"})
                    else:
                        return JsonResponse({"result": "bad",
                                             "mensaje": u"No puede modificar la fecha la entrevista ya ha sido tomada."})
                else:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al guadar los datos."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        if action == 'addlugar':
            try:
                if 'id' in request.POST:
                    integrante = IntegranteGrupoEntrevitaMsc.objects.get(status=True,
                                                                         pk=int(encrypt(request.POST['id'])),
                                                                         grupoentrevista_id=int(
                                                                             encrypt(request.POST['idg'])))
                    if not integrante.tiene_entrevista():
                        integrante.lugar = request.POST['lugar']
                        integrante.save(request)
                        return JsonResponse({"result": "ok"})
                    else:
                        return JsonResponse({"result": "bad",
                                             "mensaje": u"No puede modificar la lugar la entrevista ya ha sido tomada."})
                else:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al guadar los datos."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        if action == 'addtabalaponderacion':
            try:
                form = TablaEntrevistaMaestriaForm(request.POST)
                if form.is_valid():
                    tabla = TablaEntrevistaMaestria(nombre=form.cleaned_data['nombre'])
                    tabla.save(request)
                    for lista in json.loads(request.POST['lista_items1']):
                        estado = EstadoEntrevista(tablaentrevista=tabla, observacion=lista['nombre'],
                                                  ponderacion=float(lista['ponderacion']))
                        estado.save(request)
                    return JsonResponse({"result": "ok"})
                else:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al guadar los datos."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        if action == 'edittablapronderacion':
            try:
                form = TablaEntrevistaMaestriaForm(request.POST)
                if form.is_valid():
                    tabla = TablaEntrevistaMaestria.objects.get(status=True, pk=int(encrypt(request.POST['id'])))
                    tabla.nombre = form.cleaned_data['nombre']
                    tabla.save(request)
                    listaexclude = []
                    for lista in json.loads(request.POST['lista_items1']):
                        if int(lista['idd']) > 0:
                            listaexclude.append(int(lista['idd']))
                    estados = tabla.estadoentrevista_set.filter(status=True).exclude(pk__in=listaexclude)
                    estados.status = False
                    estados.save()
                    for lista in json.loads(request.POST['lista_items1']):
                        if EstadoEntrevista.objects.filter(status=True, tablaentrevista=tabla,
                                                           pk=lista['idd']).exists() and int(lista['idd']) > 0:
                            estado = EstadoEntrevista.objects.get(status=True, tablaentrevista=tabla, pk=lista['idd'])
                            estado.observacion = lista['nombre']
                            estado.ponderacion = float(lista['ponderacion'])
                        else:
                            estado = EstadoEntrevista(tablaentrevista=tabla, observacion=lista['nombre'],
                                                      ponderacion=float(lista['ponderacion']))
                        estado.save(request)
                    return JsonResponse({"result": "ok"})
                else:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al guadar los datos."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        if action == 'consultadetalletabla':
            try:
                if 'id' in request.POST:
                    lista = []
                    tabla = TablaEntrevistaMaestria.objects.get(pk=int(request.POST['id']))
                    for detalle in tabla.mi_detalle():
                        lista.append([detalle.id, detalle.observacion, detalle.ponderacion])
                    return JsonResponse({"result": "ok", 'lista': lista})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        if action == 'deldetalletabla':
            try:
                if 'id' in request.POST:
                    estado = EstadoEntrevista.objects.get(pk=int(encrypt(request.POST['id'])))
                    if not estado.esta_uso():
                        estado.status = False
                        estado.save()
                        return JsonResponse({"result": "ok"})
                    else:
                        return JsonResponse({"result": "bad", "mensaje": u"No puede eliminar el registro esta activo."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        if action == 'deltabla':
            try:
                if 'id' in request.POST:
                    tabla = TablaEntrevistaMaestria.objects.get(pk=int(encrypt(request.POST['id'])))
                    if not tabla.mi_detalle():
                        tabla.status = False
                        tabla.save()
                        return JsonResponse({"result": "ok"})
                    else:
                        return JsonResponse({"result": "bad", "mensaje": u"No puede eliminar el registro esta activo."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        if action == 'seleccionartabla':
            try:
                if 'id' in request.POST and 'idc' in request.POST:
                    lista = []
                    cohorte = CohorteMaestria.objects.get(status=True, pk=int(request.POST['idc']))
                    cohorte.tablaentrevista_id = int(request.POST['id'])
                    cohorte.save(request)
                    for tabla in TablaEntrevistaMaestria.objects.filter(status=True).exclude(
                            pk=int(request.POST['id'])):
                        lista.append([tabla.id])
                    return JsonResponse({"result": "ok", "idt": cohorte.tablaentrevista.id, "lista": lista})
                else:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        # REPORTES
        if action == 'pdfinscritosgrupos':
            try:
                data = {}
                data['fechaactual'] = datetime.now()
                data['grupoentrevista'] = grupoentrevista = GrupoEntrevistaMsc.objects.get(pk=int(request.POST['idg']))
                data['integrantes'] = grupoentrevista.integrantegrupoentrevitamsc_set.filter(status=True,
                                                                                             inscripcion__status=True).order_by(
                    'inscripcion__inscripcionaspirante__persona__apellido1',
                    'inscripcion__inscripcionaspirante__persona__apellido2')
                return conviert_html_to_pdf(
                    'adm_admision/pdfinscritosgrupos.html',
                    {
                        'pagesize': 'A4',
                        'data': data,
                    }
                )
            except Exception as ex:
                pass

        if action == 'pdfcertificadomatricula':
            try:
                data = {}
                data['fechaactual'] = datetime.now()
                data['inscripcioncohorte'] = inscripcioncohorte = InscripcionCohorte.objects.get(
                    pk=request.POST['idintegrante'])
                data['asignaturasmallas'] = AsignaturaMalla.objects.filter(
                    malla__carrera=inscripcioncohorte.cohortes.maestriaadmision.carrera, status=True).order_by('orden')
                return conviert_html_to_pdf(
                    'adm_admision/certificadomatricula_pdf.html',
                    {
                        'pagesize': 'A4',
                        'data': data,
                    }
                )
            except Exception as ex:
                pass

        if action == 'entrevista_pdf':
            try:
                data = {}
                data['fechaactual'] = datetime.now()
                data['cohorte'] = cohorte = CohorteMaestria.objects.get(pk=request.POST['idc'])
                data['integrante'] = IntegranteGrupoEntrevitaMsc.objects.get(pk=request.POST['idi'])
                data['estados'] = EstadoEntrevista.objects.filter(tablaentrevista=cohorte.tablaentrevista,
                                                                  status=True).order_by('id')
                listatipos = cohorte.preguntamaestria_set.values_list('pregunta__tipopregunta_id').filter(status=True)
                data['tipopreguntas'] = TipoPreguntasPrograma.objects.filter(pk__in=listatipos)
                return conviert_html_to_pdf(
                    'adm_admision/formularioentrevista_pdf.html',
                    {
                        'pagesize': 'A4',
                        'data': data,
                    }
                )
            except Exception as ex:
                pass

        if action == 'actacalificacionexamen_pdf':
            try:
                data = {}
                data['fechaactual'] = datetime.now()
                data['grupo'] = grupo = GrupoExamenMsc.objects.get(pk=int(request.POST['id']))
                data['integrantes'] = grupo.integrantegrupoexamenmsc_set.filter(status=True,
                                                                                inscripcion__status=True).order_by(
                    'inscripcion__inscripcionaspirante__persona__apellido1',
                    'inscripcion__inscripcionaspirante__persona__apellido2',
                    'inscripcion__inscripcionaspirante__persona__nombres')
                return conviert_html_to_pdf(
                    'adm_admision/actacalificacionexamen_pdf.html',
                    {
                        'pagesize': 'A4',
                        'data': data,
                    }
                )
            except Exception as ex:
                pass

        elif action == 'resetear':
            try:
                inscripcion = InscripcionCohorte.objects.get(pk=request.POST['id'])
                if not inscripcion.inscripcionaspirante.persona.emailinst:
                    per = Persona.objects.get(pk=inscripcion.inscripcionaspirante.persona.id)
                    per.emailinst = per.usuario.username + '@unemi.edu.ec'
                    per.save(request)
                else:
                    per = Persona.objects.get(pk=inscripcion.inscripcionaspirante.persona.id)
                resetear_clavepostulante(per)
                log(u'Reseteo clave de inscripción ha aspirante a maestria: %s' % inscripcion, request, "add")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        # elif action == 'emailrecordatorio':
        #     try:
        #         inscripcioncohorte = InscripcionCohorte.objects.get(pk=request.POST['id'])
        #         inscripcioncohorte.envioemailrecordatorio = True
        #         inscripcioncohorte.fecha_envioemailrecordatorio = datetime.now()
        #         inscripcioncohorte.save(request)
        #         formatocorreo = \
        #             FormatoCarreraIpec.objects.filter(carrera=inscripcioncohorte.cohortes.maestriaadmision.carrera,
        #                                               status=True)[0]
        #         lista = []
        #         if inscripcioncohorte.inscripcionaspirante.persona.emailinst:
        #             lista.append(inscripcioncohorte.inscripcionaspirante.persona.emailinst)
        #         if inscripcioncohorte.inscripcionaspirante.persona.email:
        #             lista.append(inscripcioncohorte.inscripcionaspirante.persona.email)
        #         if formatocorreo.correomaestria:
        #             lista.append(formatocorreo.correomaestria)
        #         arregloemail = [23, 24, 25, 26, 27, 28]
        #         emailaleatorio = random.choice(arregloemail)
        #         asunto = u"Requisitos para admisión de " + inscripcioncohorte.cohortes.maestriaadmision.carrera.nombre
        #         coordi = inscripcioncohorte.cohortes.maestriaadmision.carrera.coordinacion_set.filter(status=True)
        #         imagen_url = CoordinacionImagenes.objects.filter(coordinacion=coordi, tipoimagen=1, tipoimagennombre=1)[0]
        #         if formatocorreo.archivo:
        #             send_html_mail(asunto, "emails/notificacion_emailrecordatorio_ipec.html",
        #                            {'sistema': 'Posgrado UNEMI', 'inscrito': inscripcioncohorte,
        #                             'formato': formatocorreo.banner, 'imagen_url': imagen_url.imagen},
        #                            lista, [], [formatocorreo.archivo],
        #                            cuenta=variable_valor('CUENTAS_CORREOS')[emailaleatorio])
        #         else:
        #             send_html_mail(asunto, "emails/notificacion_emailrecordatorio_ipec.html",
        #                            {'sistema': 'Posgrado UNEMI', 'inscrito': inscripcioncohorte,
        #                             'formato': formatocorreo.banner, 'imagen_url': imagen_url.imagen},
        #                            lista, [], cuenta=variable_valor('CUENTAS_CORREOS')[emailaleatorio])
        #         return JsonResponse({"result": "ok"})
        #     except Exception as ex:
        #         return JsonResponse({"result": "bad", "mensaje": "%s" % ex})

        # elif action == 'emailrecordatorio_masivo':
        #     try:
        #         datos = json.loads(request.POST['lista_items1'])
        #         for dato in datos:
        #             inscripcioncohorte = InscripcionCohorte.objects.get(pk=dato['id'])
        #             inscripcioncohorte.envioemailrecordatorio = True
        #             inscripcioncohorte.fecha_envioemailrecordatorio = datetime.now()
        #             inscripcioncohorte.save(request)
        #             formatocorreo = \
        #                 FormatoCarreraIpec.objects.filter(carrera=inscripcioncohorte.cohortes.maestriaadmision.carrera,
        #                                                   status=True)[0]
        #             lista = []
        #             if inscripcioncohorte.inscripcionaspirante.persona.emailinst:
        #                 lista.append(inscripcioncohorte.inscripcionaspirante.persona.emailinst)
        #             if inscripcioncohorte.inscripcionaspirante.persona.email:
        #                 lista.append(inscripcioncohorte.inscripcionaspirante.persona.email)
        #             if formatocorreo.correomaestria:
        #                 lista.append(formatocorreo.correomaestria)
        #             arregloemail = [23, 24, 25, 26, 27, 28]
        #             emailaleatorio = random.choice(arregloemail)
        #             asunto = u"Requisitos para admisión de " + inscripcioncohorte.cohortes.maestriaadmision.carrera.nombre
        #             coordi = inscripcioncohorte.cohortes.maestriaadmision.carrera.coordinacion_set.filter(status=True)
        #             imagen_url = CoordinacionImagenes.objects.filter(coordinacion=coordi, tipoimagen=1, tipoimagennombre=1)[0]
        #             if formatocorreo.archivo:
        #                 send_html_mail(asunto, "emails/notificacion_emailrecordatorio_ipec.html",
        #                                {'sistema': 'Posgrado UNEMI', 'inscrito': inscripcioncohorte,
        #                                 'formato': formatocorreo.banner, 'imagen_url': imagen_url.imagen},
        #                                lista, [], [formatocorreo.archivo],
        #                                cuenta=variable_valor('CUENTAS_CORREOS')[emailaleatorio])
        #             else:
        #                 send_html_mail(asunto, "emails/notificacion_emailrecordatorio_ipec.html",
        #                                {'sistema': 'Posgrado UNEMI', 'inscrito': inscripcioncohorte,
        #                                 'formato': formatocorreo.banner, 'imagen_url': imagen_url.imagen},
        #                                lista, [],
        #                                cuenta=variable_valor('CUENTAS_CORREOS')[emailaleatorio])
        #             time.sleep(5)
        #         return JsonResponse({"result": "ok"})
        #     except Exception as ex:
        #         return JsonResponse({"result": "bad", "mensaje": "%s" % ex})

        elif action == 'regresaproceso':
            try:
                inscripcioncohorte = InscripcionCohorte.objects.get(pk=request.POST['id'])
                inscripcioncohorte.estado_emailevidencia = 1
                inscripcioncohorte.estado_aprobador = 1
                inscripcioncohorte.save(request)
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": "%s" % ex})

        elif action == 'descargarlistadopdf':
            try:
                data = {}
                data['fechaactual'] = datetime.now()
                data['cohorte'] = cohorte = CohorteMaestria.objects.get(pk=int(encrypt(request.POST['idcohorte'])))
                data['integrantesaprobados'] = cohorte.inscripcioncohorte_set.filter(
                    estado_aprobador=request.POST['tipoestado'], status=True).order_by(
                    'inscripcionaspirante__persona__apellido1', 'inscripcionaspirante__persona__apellido2',
                    'inscripcionaspirante__persona__nombres')
                return conviert_html_to_pdf(
                    'adm_admision/descargarlistadopdf.html',
                    {
                        'pagesize': 'A4',
                        'data': data,
                    }
                )
            except Exception as ex:
                pass

        elif action == 'actualizar_estudiantes_moodle':
            try:
                grupo = GrupoExamenMsc.objects.get(pk=6)
                grupo.crear_actualizar_estudiantes_cursogrupoex(moodle)
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'generarmatricula':
            try:
                integrante = IntegranteGrupoEntrevitaMsc.objects.get(pk=request.POST['id'])
                if integrante.inscripcion.cohortes.valormatricula:
                    valormatricula = integrante.inscripcion.cohortes.valormatricula
                    tiporubroarancel = TipoOtroRubro.objects.get(pk=2845)
                    rubro = Rubro(tipo=tiporubroarancel,
                                  persona=integrante.inscripcion.inscripcionaspirante.persona,
                                  cohortemaestria=integrante.inscripcion.cohortes,
                                  inscripcion=integrante.inscripcion,
                                  relacionados=None,
                                  nombre=tiporubroarancel.nombre + ' - ' + integrante.inscripcion.cohortes.maestriaadmision.descripcion + ' - ' + integrante.inscripcion.cohortes.descripcion,
                                  cuota=1,
                                  fecha=datetime.now().date(),
                                  fechavence=datetime.now().date() + timedelta(days=5),
                                  valor=valormatricula,
                                  iva_id=1,
                                  valoriva=0,
                                  valortotal=valormatricula,
                                  saldo=valormatricula,
                                  epunemi=True,
                                  idrubroepunemi=0,
                                  admisionposgradotipo=2,
                                  cancelado=False)
                    rubro.save(request)
                log(u'Genero rubro por concepto matricula : %s' % (integrante.inscripcion), request, "add")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'generarmatriculamasivo':
            try:
                cohorte = CohorteMaestria.objects.get(pk=int(encrypt(request.POST['id'])))
                # cohorte.procesoabierto = False
                # cohorte.save(request)
                if cohorte.valormatricula:
                    valormatricula = cohorte.valormatricula
                    integrantes = IntegranteGrupoEntrevitaMsc.objects.filter(estado_emailadmitido=2,
                                                                             cohorteadmitidasinproceso__isnull=True,
                                                                             grupoentrevista__cohortes=cohorte,
                                                                             status=True,
                                                                             inscripcion__status=True).order_by(
                        'inscripcion__inscripcionaspirante__persona__apellido1',
                        'inscripcion__inscripcionaspirante__persona__apellido2')
                    for integrante in integrantes:
                        # if integrante.inscripcion.tipocobro == 1:
                        if not integrante.inscripcion.genero_rubro_matricula():
                            tiporubroarancel = TipoOtroRubro.objects.get(pk=2845)
                            rubro = Rubro(tipo=tiporubroarancel,
                                          persona=integrante.inscripcion.inscripcionaspirante.persona,
                                          cohortemaestria=integrante.inscripcion.cohortes,
                                          inscripcion=integrante.inscripcion,
                                          relacionados=None,
                                          nombre=tiporubroarancel.nombre + ' - ' + integrante.inscripcion.cohortes.maestriaadmision.descripcion + ' - ' + integrante.inscripcion.cohortes.descripcion,
                                          cuota=1,
                                          fecha=datetime.now().date(),
                                          fechavence=integrante.inscripcion.cohortes.fechavencerubro,
                                          valor=valormatricula,
                                          iva_id=1,
                                          valoriva=0,
                                          valortotal=valormatricula,
                                          saldo=valormatricula,
                                          epunemi=True,
                                          idrubroepunemi=0,
                                          admisionposgradotipo=2,
                                          cancelado=False)
                            rubro.save(request)
                            integrante.inscripcion.tipocobro = 2
                            integrante.inscripcion.tipo_id = 2845
                            integrante.inscripcion.save(request)
                            log(u'Genero rubro por concepto matricula : %s' % (integrante.inscripcion), request,
                                "add")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'generarcohortematriculamasivo':
            try:
                cohorte = CohorteMaestria.objects.get(pk=int(encrypt(request.POST['id'])))
                if cohorte.valormatricula:
                    valormatricula = cohorte.valormatricula
                    if cohorte.tipo == 2:
                        integrantes = InscripcionCohorte.objects.filter(cohortes=cohorte, cohortes__tipo=2,
                                                                        estado_aprobador=2,
                                                                        integrantegrupoexamenmsc__estado=2, status=True)
                    if cohorte.tipo == 3:
                        integrantes = InscripcionCohorte.objects.filter(cohortes=cohorte, estado_aprobador=2,
                                                                        status=True)
                    for integrante in integrantes:
                        # if integrante.tipocobro == 1:
                        if not integrante.genero_rubro_matricula():
                            tiporubroarancel = TipoOtroRubro.objects.get(pk=2845)
                            rubro = Rubro(tipo=tiporubroarancel,
                                          persona=integrante.inscripcionaspirante.persona,
                                          cohortemaestria=integrante.cohortes,
                                          inscripcion=integrante,
                                          relacionados=None,
                                          nombre=tiporubroarancel.nombre + ' - ' + integrante.cohortes.maestriaadmision.descripcion + ' - ' + integrante.cohortes.descripcion,
                                          cuota=1,
                                          fecha=datetime.now().date(),
                                          fechavence=integrante.cohortes.fechavencerubro,
                                          valor=valormatricula,
                                          iva_id=1,
                                          valoriva=0,
                                          valortotal=valormatricula,
                                          saldo=valormatricula,
                                          epunemi=True,
                                          idrubroepunemi=0,
                                          admisionposgradotipo=2,
                                          cancelado=False)
                            rubro.save(request)
                            integrante.tipocobro = 2
                            integrante.tipo_id = 2845
                            integrante.save(request)
                            log(u'Genero rubro por concepto matricula : %s' % (integrante), request, "add")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'generarcohortematriculaindividual':
            try:
                cohorte = CohorteMaestria.objects.get(pk=int(encrypt(request.POST['id'])))
                if cohorte.valormatricula:
                    valormatricula = cohorte.valormatricula
                    integrante = InscripcionCohorte.objects.filter(pk=request.POST['insc'], status=True).last()
                    if not integrante.genero_rubro_matricula():
                        tiporubroarancel = TipoOtroRubro.objects.get(pk=2845)
                        rubro = Rubro(tipo=tiporubroarancel,
                                      persona=integrante.inscripcionaspirante.persona,
                                      cohortemaestria=integrante.cohortes,
                                      inscripcion=integrante,
                                      relacionados=None,
                                      nombre=tiporubroarancel.nombre + ' - ' + integrante.cohortes.maestriaadmision.descripcion + ' - ' + integrante.cohortes.descripcion,
                                      cuota=1,
                                      fecha=datetime.now().date(),
                                      fechavence=integrante.cohortes.fechavencerubro,
                                      valor=valormatricula,
                                      iva_id=1,
                                      valoriva=0,
                                      valortotal=valormatricula,
                                      saldo=valormatricula,
                                      epunemi=True,
                                      idrubroepunemi=0,
                                      admisionposgradotipo=2,
                                      cancelado=False)
                        rubro.save(request)
                        integrante.tipocobro = 2
                        integrante.tipo_id = 2845
                        integrante.save(request)
                        log(u'Genero rubro por concepto matrícula : %s' % (integrante), request, "add")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'generarcohorteprogramamasivo':
            try:
                cohorte = CohorteMaestria.objects.get(pk=int(encrypt(request.POST['id'])))
                if cohorte.valorprograma:
                    valorprograma = cohorte.valorprograma
                    if cohorte.tipo == 2:
                        integrantes = InscripcionCohorte.objects.filter(cohortes=cohorte, cohortes__tipo=2,
                                                                        estado_aprobador=2,
                                                                        integrantegrupoexamenmsc__estado=2, status=True)
                    if cohorte.tipo == 3:
                        integrantes = InscripcionCohorte.objects.filter(cohortes=cohorte, estado_aprobador=2,
                                                                        status=True)
                    for integrante in integrantes:
                        # if integrante.tipocobro == 1:
                        if not integrante.genero_rubro_programa():
                            tiporubroarancel = TipoOtroRubro.objects.get(pk=integrante.cohortes.tiporubro.id)
                            rubro = Rubro(tipo=tiporubroarancel,
                                          persona=integrante.inscripcionaspirante.persona,
                                          cohortemaestria=integrante.cohortes,
                                          inscripcion=integrante,
                                          relacionados=None,
                                          nombre=tiporubroarancel.nombre + ' - ' + integrante.cohortes.maestriaadmision.descripcion + ' - ' + integrante.cohortes.descripcion,
                                          cuota=1,
                                          fecha=datetime.now().date(),
                                          fechavence=integrante.cohortes.fechavencerubro,
                                          valor=valorprograma,
                                          iva_id=1,
                                          valoriva=0,
                                          valortotal=valorprograma,
                                          saldo=valorprograma,
                                          epunemi=True,
                                          idrubroepunemi=0,
                                          admisionposgradotipo=3,
                                          cancelado=False)
                            rubro.save(request)
                            integrante.tipocobro = 3
                            integrante.tipo = tiporubroarancel
                            integrante.save(request)
                            log(u'Genero rubro por concepto programa maestría : %s' % (integrante), request, "add")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'generarcohorteprogramaindividual':
            try:
                cohorte = CohorteMaestria.objects.get(pk=int(encrypt(request.POST['id'])))
                if cohorte.valorprograma:
                    valorprograma = cohorte.valorprograma
                    integrante = InscripcionCohorte.objects.filter(pk=request.POST['insc'], status=True).last()
                    if not integrante.genero_rubro_programa():
                        tiporubroarancel = TipoOtroRubro.objects.get(pk=integrante.cohortes.tiporubro.id)
                        rubro = Rubro(tipo=tiporubroarancel,
                                      persona=integrante.inscripcionaspirante.persona,
                                      cohortemaestria=integrante.cohortes,
                                      inscripcion=integrante,
                                      relacionados=None,
                                      nombre=tiporubroarancel.nombre + ' - ' + integrante.cohortes.maestriaadmision.descripcion + ' - ' + integrante.cohortes.descripcion,
                                      cuota=1,
                                      fecha=datetime.now().date(),
                                      fechavence=integrante.cohortes.fechavencerubro,
                                      valor=valorprograma,
                                      iva_id=1,
                                      valoriva=0,
                                      valortotal=valorprograma,
                                      saldo=valorprograma,
                                      epunemi=True,
                                      idrubroepunemi=0,
                                      admisionposgradotipo=3,
                                      cancelado=False)
                        rubro.save(request)
                        integrante.tipocobro = 3
                        integrante.tipo = tiporubroarancel
                        integrante.save(request)
                        log(u'Genero rubro por concepto programa maestría : %s' % (integrante), request, "add")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'generarprogramamasivo':
            try:
                cohorte = CohorteMaestria.objects.get(pk=int(encrypt(request.POST['id'])))
                # cohorte.procesoabierto = False
                # cohorte.save(request)
                if cohorte.valorprograma:
                    valorprograma = cohorte.valorprograma
                    integrantes = IntegranteGrupoEntrevitaMsc.objects.filter(estado_emailadmitido=2,
                                                                             cohorteadmitidasinproceso__isnull=True,
                                                                             grupoentrevista__cohortes=cohorte,
                                                                             status=True,
                                                                             inscripcion__status=True).order_by(
                        'inscripcion__inscripcionaspirante__persona__apellido1',
                        'inscripcion__inscripcionaspirante__persona__apellido2')
                    for integrante in integrantes:
                        # if integrante.inscripcion.tipocobro == 1:
                        if not integrante.inscripcion.genero_rubro_programa():
                            tiporubroarancel = TipoOtroRubro.objects.get(
                                pk=integrante.inscripcion.cohortes.tiporubro.id)
                            rubro = Rubro(tipo=tiporubroarancel,
                                          persona=integrante.inscripcion.inscripcionaspirante.persona,
                                          cohortemaestria=integrante.inscripcion.cohortes,
                                          inscripcion=integrante.inscripcion,
                                          relacionados=None,
                                          nombre=tiporubroarancel.nombre + ' - ' + integrante.inscripcion.cohortes.maestriaadmision.descripcion + ' - ' + integrante.inscripcion.cohortes.descripcion,
                                          cuota=1,
                                          fecha=datetime.now().date(),
                                          fechavence=integrante.inscripcion.cohortes.fechavencerubro,
                                          valor=valorprograma,
                                          iva_id=1,
                                          valoriva=0,
                                          valortotal=valorprograma,
                                          saldo=valorprograma,
                                          epunemi=True,
                                          idrubroepunemi=0,
                                          admisionposgradotipo=3,
                                          cancelado=False)
                            rubro.save(request)
                            integrante.inscripcion.tipocobro = 3
                            integrante.inscripcion.tipo = tiporubroarancel
                            integrante.inscripcion.save(request)
                            log(u'Genero rubro por concepto programa maestría : %s' % (integrante.inscripcion),
                                request, "add")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'importarinscritos':
            try:
                idgrupoexamen = request.POST['idgrupoexamen']
                lista = request.POST['lista'].split(',')
                for elemento in lista:
                    if not IntegranteGrupoExamenMsc.objects.filter(inscripcion__status=True, inscripcion_id=elemento,
                                                                   grupoexamen_id=idgrupoexamen, status=True):
                        integrantes = IntegranteGrupoExamenMsc(inscripcion_id=elemento,
                                                               grupoexamen_id=idgrupoexamen)
                        integrantes.save(request)
                return JsonResponse({'result': 'ok'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'importarrequisitos':
            try:
                idcohorte = int(encrypt(request.POST['idcohorte']))
                lista = request.POST['lista'].split(',')
                for elemento in lista:
                    if not RequisitosMaestria.objects.filter(requisito_id=elemento, cohorte_id=idcohorte, status=True):
                        requisitomaestria = RequisitosMaestria(requisito_id=elemento,
                                                               cohorte_id=idcohorte)
                        requisitomaestria.save(request)
                return JsonResponse({'result': 'ok'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'importarinscritosentrevista':
            try:
                idgrupoentrevista = request.POST['idgrupoentrevista']
                lista = request.POST['lista'].split(',')
                for elemento in lista:
                    if not IntegranteGrupoEntrevitaMsc.objects.filter(inscripcion__status=True, inscripcion_id=elemento,
                                                                      grupoentrevista_id=idgrupoentrevista,
                                                                      status=True):
                        integrantesentrevista = IntegranteGrupoEntrevitaMsc(inscripcion_id=elemento,
                                                                            grupoentrevista_id=idgrupoentrevista)
                        integrantesentrevista.save(request)
                return JsonResponse({'result': 'ok'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'modificarnotaexamen':
            try:
                integrante = IntegranteGrupoExamenMsc.objects.get(pk=request.POST['mid'])
                notapasarexamen = int(integrante.inscripcion.cohortes.notaminimaexa)
                valor = request.POST['vc']
                integrante.notafinal = float(valor)
                if float(valor) >= float(notapasarexamen):
                    integrante.estado = 2
                if float(valor) < float(notapasarexamen):
                    integrante.estado = 3
                integrante.save(request)
                return JsonResponse({'result': 'ok', 'valor': integrante.notafinal})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', "mensaje": u"Error al actualizar nota."})

        elif action == 'guardarobservacion':
            try:
                cohorte = CohorteMaestria.objects.get(pk=int(encrypt(request.POST['id'])))
                cohorte.observacionrubro = request.POST['observacion']
                cohorte.save(request)
                return JsonResponse({"result": False}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'modificarnotaentrevista':
            try:
                integranteentrevista = IntegranteGrupoEntrevitaMsc.objects.get(pk=request.POST['mid'])
                valor = request.POST['vc']
                notaaprobar = int(integranteentrevista.inscripcion.cohortes.notaminimaexa)
                notalfinalexamen = \
                    integranteentrevista.inscripcion.integrantegrupoexamenmsc_set.filter(inscripcion__status=True,
                                                                                         status=True)[0].notafinal
                integranteentrevista.notaentrevista = 0
                integranteentrevista.entrevista = False
                integranteentrevista.estado_emailadmitido = 1
                if valor == 'y':
                    integranteentrevista.estado_emailadmitido = 2
                    integranteentrevista.entrevista = True
                    integranteentrevista.notaentrevista = 100
                notalfinalenttrevista = notalfinalexamen + integranteentrevista.notaentrevista
                integranteentrevista.notafinal = round(notalfinalenttrevista, 0)
                if integranteentrevista.entrevista:
                    integranteentrevista.estado = 2
                else:
                    integranteentrevista.estado = 3
                integranteentrevista.save(request)
                return JsonResponse({'result': 'ok'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', "mensaje": u"Error al actualizar nota."})
            #     integranteentrevista = IntegranteGrupoEntrevitaMsc.objects.get(pk=request.POST['mid'])
            #     valor = int(request.POST['vc'])
            #     notaaprobar = int(integranteentrevista.inscripcion.cohortes.notaminimaexa)
            #     notalfinalexamen = integranteentrevista.inscripcion.integrantegrupoexamenmsc_set.filter(status=True)[0].notafinal
            #     integranteentrevista.notaentrevista = valor
            #     notalfinalenttrevista = notalfinalexamen + valor
            #     integranteentrevista.notafinal = round(notalfinalenttrevista, 0)
            #     if round(notalfinalenttrevista, 0) >= notaaprobar:
            #         integranteentrevista.estado = 2
            #     if round(notalfinalenttrevista, 0) < notaaprobar:
            #         integranteentrevista.estado = 3
            #     integranteentrevista.save(request)
            #     return JsonResponse({'result': 'ok', 'valor': integranteentrevista.notaentrevista})
            # except Exception as ex:
            #     transaction.set_rollback(True)
            #     return JsonResponse({'result': 'bad',"mensaje": u"Error al actualizar nota."})

        elif action == 'changeesobligatorio':
            try:
                requisitomaestria = RequisitosMaestria.objects.get(pk=request.POST['idreq'])
                if requisitomaestria.obligatorio:
                    requisitomaestria.obligatorio = False
                else:
                    requisitomaestria.obligatorio = True
                requisitomaestria.save(request)
                return JsonResponse({'result': 'ok', 'valor': requisitomaestria.obligatorio})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        if action == 'aprobarrequisitos':
            try:
                cohorte = CohorteMaestria.objects.get(status=True, pk=int(request.POST['idcohorte']))
                inscripcioncohorte = InscripcionCohorte.objects.get(pk=int(request.POST['id']), status=True)
                for re in cohorte.requisitosmaestria_set.filter(status=True):
                    ingresoevidencias = re.detalle_requisitosmaestriacohorte(inscripcioncohorte)
                    if ingresoevidencias:
                        if ingresoevidencias.ultima_evidencia().estado_aprobacion == 1:
                            detalleevidencia = ingresoevidencias.ultima_evidencia()
                            detalleevidencia.observacion_aprobacion = 'APROBADO CORRECTAMENTE'
                            detalleevidencia.persona = persona
                            detalleevidencia.estado_aprobacion = 2
                            detalleevidencia.fecha_aprobacion = datetime.now()
                            detalleevidencia.save(request)
                            log(u'Actualizó observacion evidencia masivo: %s' % (detalleevidencia), request, "add")
                # verificar si ya esta todo aprobado para enviar correo de aprobacion
                bandera = 0
                for re in cohorte.requisitosmaestria_set.filter(status=True, obligatorio=True):
                    ingresoevidencias = re.detalle_requisitosmaestriacohorte(inscripcioncohorte)
                    if not ingresoevidencias.ultima_evidencia().estado_aprobacion == 2:
                        bandera = 1
                if bandera == 0:
                    # inscripcioncoohorte = InscripcionCohorte.objects.get(pk=request.POST['idinscripcioncohorte'])
                    inscripcioncohorte.fecha_emailevidencia = datetime.now()
                    inscripcioncohorte.estado_emailevidencia = 2
                    inscripcioncohorte.persona_emailevidencia = persona
                    inscripcioncohorte.fecha_aprobador = datetime.now()
                    inscripcioncohorte.estado_aprobador = 2
                    inscripcioncohorte.persona_aprobador = persona
                    inscripcioncohorte.save(request)
                    log(u'Envio email aprobacion, masivo: %s' % (inscripcioncohorte), request, "add")
                    send_html_mail("Aprobado Admision-UNEMI.", "emails/registroaprobadomasivo.html",
                                   {'sistema': u'Admision - UNEMI', 'fecha': datetime.now().date(),
                                    'inscripcioncohorte': inscripcioncohorte,
                                    'hora': datetime.now().time(), 't': miinstitucion()},
                                   inscripcioncohorte.inscripcionaspirante.persona.emailpersonal(), [],
                                   cuenta=variable_valor('CUENTAS_CORREOS')[16])
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u'Error al guardar los datos'})

        elif action == 'extraernotasmoodle':
            try:
                grupo = GrupoExamenMsc.objects.get(pk=request.POST['id'], status=True)
                for alumno in grupo.mis_integrantes():
                    # Extraer datos de moodle
                    for notasmooc in grupo.notas_de_moodle(alumno.inscripcion.inscripcionaspirante.persona):
                        campo = alumno.campo(notasmooc[1].upper())
                        if type(notasmooc[0]) is Decimal:
                            if null_to_decimal(campo) != float(notasmooc[0]):
                                actualizar_nota_grupo(request, alumno.id, notasmooc[1].upper(), notasmooc[0])
                            else:
                                if alumno.notafinal:
                                    if alumno.notafinal >= grupo.cohorte.notaminimaexa:
                                        alumno.estado = 2
                                    if alumno.notafinal < grupo.cohorte.notaminimaexa:
                                        alumno.estado = 3
                                else:
                                    alumno.estado = 3
                                alumno.save(request)
                        else:
                            if null_to_decimal(campo) != float(0):
                                actualizar_nota_grupo(request, alumno.id, notasmooc[1].upper(), notasmooc[0])
                log(u'Importó notas de moodle al grupo de examen: %s' % (grupo), request, "add")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al actualizar los datos."})

        elif action == 'copiaraentrevista':
            try:
                grupo = GrupoExamenMsc.objects.get(pk=request.POST['id'], status=True)
                grupoentrevista = GrupoEntrevistaMsc.objects.get(pk=request.POST['idd'], status=True)
                alumnos_grupo = grupo.mis_integrantes()
                alumnos_entrevista = grupoentrevista.mis_integrantes()

                # Obtener las notas finales de los alumnos de examen correspondientes a los de la entrevista
                notas_finales = [(alumnoexam.notaexa, alumnoexam.notafinal, alumnoexam.estado) for alumnoexam in alumnos_grupo
                                 if alumnoexam.inscripcion in [alumnoentrev.inscripcion for alumnoentrev in
                                                               alumnos_entrevista]]

                # Copiar las notas finales a los alumnos de la entrevista correspondientes
                for i, alumnoentrev in enumerate(alumnos_entrevista):
                    if alumnoentrev.inscripcion in [alumno.inscripcion for alumno in alumnos_grupo]:
                        if notas_finales[i][0] is not None and notas_finales[i][1] is not None:
                            copiar_nota_entrevista(alumnoentrev.id, notas_finales[i][0], notas_finales[i][1], notas_finales[i][2])

                # Crear objetos de integrante para los alumnos que no estén en la entrevista
                integrantes_entrevista = [IntegranteGrupoEntrevitaMsc(inscripcion_id=alumno.inscripcion.id,
                                                                      grupoentrevista_id=grupoentrevista.id,
                                                                      notaentrevista=alumno.notaexa,
                                                                      notafinal=round((alumno.notaexa + alumno.notafinal)/2,2),
                                                                      estado=alumno.estado,
                                                                      fecha_creacion=datetime.now(),
                                                                      usuario_creacion = persona.usuario
                                                                      )
                                          for alumno in alumnos_grupo
                                          if alumno.notaexa is not None and alumno.notafinal is not None and not IntegranteGrupoEntrevitaMsc.objects.filter(status=True, inscripcion=alumno.inscripcion).exists() and alumno.inscripcion
                                          not in [alumnoentrev.inscripcion for alumnoentrev in alumnos_entrevista]]

                IntegranteGrupoEntrevitaMsc.objects.bulk_create(
                    integrantes_entrevista)  # Crear los objetos de integrante en la BD
                log(u'Actualizó notas de entrevista mediante copiar: %s' % (alumnos_entrevista), request, "edit")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al copiar los datos."})


        elif action == 'cargararchivo':
            try:
                if 'archivo' in request.FILES:
                    newfile = request.FILES['archivo']
                    if newfile.size > 10485760:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Error, Tamaño de archivo Maximo permitido es de 10Mb"})
                    else:
                        newfiles = request.FILES['archivo']
                        newfilesd = newfiles._name
                        ext = newfilesd[newfilesd.rfind("."):]
                        if not ext.lower() == '.pdf':
                            return JsonResponse({"result": "bad", "mensaje": u"Error, solo archivos .pdf."})
                f = RequisitosMaestriaForm(request.POST)
                if f.is_valid():
                    inscripcioncohorte = InscripcionCohorte.objects.get(pk=request.POST['id'])
                    requisitosmaestria = RequisitosMaestria.objects.get(pk=request.POST['idevidencia'])
                    if not EvidenciaRequisitosAspirante.objects.filter(requisitos=requisitosmaestria,
                                                                       inscripcioncohorte=inscripcioncohorte,
                                                                       status=True).exists():
                        requisitomaestria = EvidenciaRequisitosAspirante(requisitos=requisitosmaestria,
                                                                         inscripcioncohorte=inscripcioncohorte)
                        requisitomaestria.save(request)
                        if 'archivo' in request.FILES:
                            newfile = request.FILES['archivo']
                            newfile._name = generar_nombre("requisitopgrado_", newfile._name)
                            requisitomaestria.archivo = newfile
                            requisitomaestria.save(request)
                        log(u'Adicionó requisito de maestria aspirante: %s' % requisitomaestria.requisitos, request,
                            "add")
                        detalle = DetalleEvidenciaRequisitosAspirante(evidencia=requisitomaestria,
                                                                      estadorevision=1,
                                                                      fecha=datetime.now().date(),
                                                                      observacion=f.cleaned_data['observacion'])
                        detalle.save(request)
                        return JsonResponse({"result": "ok"})
                    else:
                        if 'archivo' in request.FILES:
                            newfile = request.FILES['archivo']
                            if newfile.size > 10485760:
                                return JsonResponse({"result": "bad",
                                                     "mensaje": u"Error, Tamaño de archivo Maximo permitido es de 10Mb"})
                            else:
                                newfiles = request.FILES['archivo']
                                newfilesd = newfiles._name
                                ext = newfilesd[newfilesd.rfind("."):]
                                if not ext.lower() == '.pdf':
                                    return JsonResponse({"result": "bad", "mensaje": u"Error, solo archivos .pdf."})
                        f = RequisitosMaestriaForm(request.POST)
                        requisito = EvidenciaRequisitosAspirante.objects.get(requisitos=requisitosmaestria,
                                                                             inscripcioncohorte=inscripcioncohorte,
                                                                             status=True)
                        if f.is_valid():
                            if 'archivo' in request.FILES:
                                newfile = request.FILES['archivo']
                                newfile._name = generar_nombre("requisitopgrado_", newfile._name)
                                requisito.archivo = newfile
                                requisito.save(request)
                            detalle = DetalleEvidenciaRequisitosAspirante(evidencia=requisito,
                                                                          estadorevision=1,
                                                                          # persona=inscripcioncohorte.inscripcionaspirante.persona,
                                                                          fecha=datetime.now().date(),
                                                                          observacion=f.cleaned_data['observacion'])
                            detalle.save(request)
                            log(u'Editó requisito de maestría aspirante: %s' % requisito, request, "edit")
                            return JsonResponse({"result": "ok"})
                        else:
                            raise NameError('Error')
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'cargararchivoimagen':
            try:
                if 'archivo' in request.FILES:
                    newfile = request.FILES['archivo']
                    if newfile.size > 10485760:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Error, Tamaño de archivo Maximo permitido es de 10Mb"})
                    else:
                        newfiles = request.FILES['archivo']
                        newfilesd = newfiles._name
                        ext = newfilesd[newfilesd.rfind("."):]
                        if ext.lower() == '.jpg' or ext.lower() == '.png' or ext.lower() == '.jpeg' or ext.lower() == '.JPG' or ext.lower() == '.PNG' or ext.lower() == '.JPEG':
                            a = 1
                        else:
                            return JsonResponse({"result": "bad", "mensaje": u"Error, solo archivos .jpg, .png."})
                f = RequisitosMaestriaImgForm(request.POST)
                if f.is_valid():
                    inscripcioncohorte = InscripcionCohorte.objects.get(pk=request.POST['id'])
                    requisitosmaestria = RequisitosMaestria.objects.get(pk=request.POST['idevidencia'])
                    if not EvidenciaRequisitosAspirante.objects.filter(requisitos=requisitosmaestria,
                                                                       inscripcioncohorte=inscripcioncohorte,
                                                                       status=True).exists():
                        requisitomaestria = EvidenciaRequisitosAspirante(requisitos=requisitosmaestria,
                                                                         inscripcioncohorte=inscripcioncohorte)
                        requisitomaestria.save(request)
                        if 'archivo' in request.FILES:
                            newfile = request.FILES['archivo']
                            newfile._name = generar_nombre("requisitopgrado_", newfile._name)
                            requisitomaestria.archivo = newfile
                            requisitomaestria.save(request)
                        log(u'Adicionó requisito de maestria aspirante: %s' % requisitomaestria.requisitos, request,
                            "add")
                        detalle = DetalleEvidenciaRequisitosAspirante(evidencia=requisitomaestria,
                                                                      estadorevision=1,
                                                                      fecha=datetime.now().date(),
                                                                      observacion=f.cleaned_data['observacion'])
                        detalle.save(request)
                        return JsonResponse({"result": "ok"})
                    else:
                        if 'archivo' in request.FILES:
                            newfile = request.FILES['archivo']
                            if newfile.size > 10485760:
                                return JsonResponse({"result": "bad",
                                                     "mensaje": u"Error, Tamaño de archivo Maximo permitido es de 10Mb"})
                            else:
                                newfiles = request.FILES['archivo']
                                newfilesd = newfiles._name
                                ext = newfilesd[newfilesd.rfind("."):]
                                if ext.lower() == '.jpg' or ext.lower() == '.png' or ext.lower() == '.jpeg' or ext.lower() == '.JPG' or ext.lower() == '.PNG' or ext.lower() == '.JPEG':
                                    a = 1
                                else:
                                    return JsonResponse(
                                        {"result": "bad", "mensaje": u"Error, solo archivos .jpg, .png."})
                        f = RequisitosMaestriaImgForm(request.POST)
                        requisito = EvidenciaRequisitosAspirante.objects.get(requisitos=requisitosmaestria,
                                                                             inscripcioncohorte=inscripcioncohorte,
                                                                             status=True)
                        if f.is_valid():
                            if 'archivo' in request.FILES:
                                newfile = request.FILES['archivo']
                                newfile._name = generar_nombre("requisitopgrado_", newfile._name)
                                requisito.archivo = newfile
                                requisito.save(request)
                            detalle = DetalleEvidenciaRequisitosAspirante(evidencia=requisito,
                                                                          estadorevision=1,
                                                                          fecha=datetime.now().date(),
                                                                          observacion=f.cleaned_data['observacion'])
                            detalle.save(request)
                            log(u'Editó requisito de maestría aspirante: %s' % requisito, request, "edit")
                            return JsonResponse({"result": "ok"})
                        else:
                            raise NameError('Error')
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'cerrarproceso':
            try:
                tipo = int(request.POST['tipo'])
                cohorte = CohorteMaestria.objects.get(pk=int(encrypt(request.POST['idcohorte'])))
                if tipo == 1:
                    cohorte.procesoabierto = False
                if tipo == 2:
                    cohorte.procesoabierto = True
                cohorte.save(request)
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'delinscripcion':
            try:
                inscripcioncohorte = InscripcionCohorte.objects.get(pk=request.POST['idinscripcion'])
                inscripcioncohorte.status = False
                inscripcioncohorte.save(request)
                log(u'Eliminó inscripcion: %s - %s' % (inscripcioncohorte, inscripcioncohorte.inscripcionaspirante),
                    request, "del")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar."})

        elif action == 'buscarprofesor':
            try:
                lista = []
                for profesor in Profesor.objects.all():
                    lista.append([profesor.id, profesor.persona.nombre_completo_inverso()])
                return JsonResponse({"result": "ok", "lista": lista})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'aprovarrequisitosmasivo':
            try:
                fecha_actual = datetime.now()
                obsr = request.POST['obser']
                estado = int(request.POST['estado'])
                id = request.POST['id']
                id_req = json.loads(request.POST['id_requisitos'])
                inscripcioncohorte = InscripcionCohorte.objects.get(pk=int(id))
                detalleevidencia = DetalleEvidenciaRequisitosAspirante.objects.filter(status=True, evidencia__inscripcioncohorte=inscripcioncohorte, evidencia__requisitos__id__in=id_req)
                evidencias_rech = ''
                for deta in detalleevidencia:
                    evidencias_rech+=f'{deta.evidencia.requisitos.requisito.nombre}, '
                    deta.estado_aprobacion = estado
                    deta.observacion_aprobacion = obsr
                    deta.fecha_aprobacion = fecha_actual
                    deta.persona = persona
                    deta.save(request)
                    log(u'%s los requisitos de manera mavisa: %s - %s' % (u'Aprobó' if estado == 2 else 'Rechazó',inscripcioncohorte, deta),
                        request, "change")
                cohorte = inscripcioncohorte.cohortes
                # verificar si ya esta todo aprobado para enviar correo de aprobacion

                claserequisitoadmision = ClaseRequisito.objects.values_list('requisito__id').filter(clasificacion=1,
                                                                                                    status=True)
                requisitosadmision = cohorte.requisitosmaestria_set.filter(requisito__in=claserequisitoadmision,
                                                                           obligatorio=True, status=True).order_by('id')
                bandera = 0
                for re in requisitosadmision:
                    ingresoevidencias = re.detalle_requisitosmaestriacohorte(inscripcioncohorte)
                    if ingresoevidencias:
                        if not ingresoevidencias.ultima_evidencia().estado_aprobacion == 2:
                            bandera = 1
                    else:
                        bandera = 1
                if bandera == 1 and estado == 3:
                    asunto = u"REQUISITO DE ADMISIÓN RECHAZADO"
                    observacion = f'Se le comunica que la Secretaría Técnica de Posgrado ha rechazado el requisito {evidencias_rech}, con la siguiente observación {obsr}. Por favor, comunicarse con el postulante {inscripcioncohorte.inscripcionaspirante.persona.nombre_completo_inverso()} para subir el requisito correctamente. Puede dar clic en la URL para ser redirigido directamente al postulante  que presenta dicha observación.'
                    para = inscripcioncohorte.asesor.persona
                    perfiu = inscripcioncohorte.asesor.perfil_administrativo()
                    notificacion3(asunto, observacion, para, None, '/comercial?s=' + inscripcioncohorte.inscripcionaspirante.persona.cedula, inscripcioncohorte.pk, 1,
                                 'sga', InscripcionCohorte, perfiu, request)

                    inscripcioncohorte.preaprobado = False
                    inscripcioncohorte.todosubido = False
                    inscripcioncohorte.tienerechazo = True

                    if inscripcioncohorte.estado_aprobador == 2:
                        inscripcioncohorte.estado_aprobador = 1
                    inscripcioncohorte.save(request)
                    log(u'Postulante con evidencia rechazada masiva: %s' % (inscripcioncohorte), request, "edit")

                    # send_html_mail("RECHAZADA EVIDENCIA - UNEMI.", "emails/registrorechazadoevidencia.html",
                    #                {'sistema': u'ADMISIÓN - UNEMI', 'fecha': datetime.now().date(),
                    #                 'hora': datetime.now().time(), 't': miinstitucion(),
                    #                 'requisito': evidencias_rech,
                    #                 'observacion': obsr},
                    #                inscripcioncohorte.inscripcionaspirante.persona.emailpersonal(),
                    #                [],
                    #                cuenta=variable_valor('CUENTAS_CORREOS')[16])
                if bandera == 0:
                    # inscripcioncoohorte = InscripcionCohorte.objects.get(pk=request.POST['idinscripcioncohorte'])
                    inscripcioncohorte.fecha_emailevidencia = datetime.now()
                    inscripcioncohorte.estado_emailevidencia = 2
                    inscripcioncohorte.persona_emailevidencia = persona
                    inscripcioncohorte.fecha_aprobador = datetime.now()
                    inscripcioncohorte.estado_aprobador = 2
                    inscripcioncohorte.persona_aprobador = persona
                    inscripcioncohorte.tienerechazo = False
                    inscripcioncohorte.save(request)

                    asunto = u"POSTULANTE ADMITIDO"
                    observacion = f'Se le comunica que el postulante {inscripcioncohorte.inscripcionaspirante.persona} con cédula {inscripcioncohorte.inscripcionaspirante.persona.cedula} ha subido todos sus requisitos de admisión. Por favor, pre-aprobar los requisitos del postulante. Puede dar clic en la URL para ser redirigido directamente al postulante pendiente de pre-aprobación.'
                    para = inscripcioncohorte.asesor.persona
                    perfiu = inscripcioncohorte.asesor.perfil_administrativo()

                    notificacion3(asunto, observacion, para, None,
                                  '/comercial?s=' + inscripcioncohorte.inscripcionaspirante.persona.cedula,
                                  inscripcioncohorte.pk, 1,
                                  'sga', InscripcionCohorte, perfiu, request)

                    log(u'Envio email aprobacion, masivo: %s' % (inscripcioncohorte), request, "add")
                    send_html_mail("Aprobado Admision-UNEMI.", "emails/registroaprobadomasivo.html",
                                   {'sistema': u'Admision - UNEMI', 'fecha': datetime.now().date(),
                                    'inscripcioncohorte': inscripcioncohorte,
                                    'hora': datetime.now().time(), 't': miinstitucion()},
                                   inscripcioncohorte.inscripcionaspirante.persona.emailpersonal(), [],
                                   cuenta=variable_valor('CUENTAS_CORREOS')[16])
                return JsonResponse({"result":True})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar."})

        elif action == 'importarrequisitosho':
            try:
                with transaction.atomic():
                    ids = request.POST['ids'].split(',')
                    idmaestria = int(request.POST['idm'])

                    for id in ids:
                        if not RequisitosMaestria.objects.filter(requisito_id=id, maestria_id=idmaestria, status=True):
                            requisitomaestria = RequisitosMaestria(requisito_id=id, maestria_id=idmaestria)
                            requisitomaestria.save(request)
                            log(u'Adicionó requisito de homologación:: %s' % requisitomaestria, request, "add")

                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'copiarrequisitos':
            try:
                with transaction.atomic():
                    f = CopiarRequisitosForm(request.POST)
                    if f.is_valid():
                        idmaestria = int(encrypt(request.POST['id']))
                        idrequisitos = RequisitosMaestria.objects.filter(status=True, maestria=f.cleaned_data['programa']).values_list('requisito__id', flat=True)

                        for id in idrequisitos:
                            if not RequisitosMaestria.objects.filter(requisito_id=id, maestria_id=idmaestria,
                                                                     status=True):
                                requisitomaestria = RequisitosMaestria(requisito_id=id, maestria_id=idmaestria)
                                requisitomaestria.save(request)
                                log(u'Copió requisitos de homologación: %s' % requisitomaestria, request, "add")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({'error': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                             "message": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        return JsonResponse({"result": "bad", "mensaje": u"Solicitud Incorrecta."})

    else:
        if 'action' in request.GET:
            data['action'] = action = request.GET['action']

            if action == 'addmaestriaadmision':
                try:
                    form = AdmiPeriodoForm()
                    maestria = MaestriasAdmision.objects.values_list('carrera_id', flat=True).filter(status=True)
                    malla = Malla.objects.values_list('carrera_id', flat=True).filter(status=True, cerrado=False,
                                                                                      vigente=True)
                    form.fields['carrera'].queryset = Carrera.objects.filter(status=True, coordinacion__id__in=[7, 13],
                                                                             id__in=malla).exclude(id__in=maestria)
                    data['form2'] = form
                    template = get_template('adm_admision/addperiodoadmision.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'cargararchivo':
                try:
                    data['title'] = u'Evidencias de requisitos de maestría'
                    data['id'] = request.GET['id']
                    data['tipoestado'] = request.GET['tipoestado']
                    if request.GET['cant_requisitos'] != '':
                        data['cant_requisitos'] = int(request.GET['cant_requisitos'])
                    else:
                        data['cant_requisitos'] = request.GET['cant_requisitos']
                    data['idevidencia'] = request.GET['idevidencia']
                    requisito = RequisitosMaestria.objects.get(pk=int(request.GET['idevidencia']), status=True)
                    form = RequisitosMaestriaForm()
                    data['form'] = form
                    template = get_template("adm_admision/add_requisitomaestria.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content,
                                         "nombre": "SUBIR DOCUMENTO DE REQUISITO " + str(requisito.requisito.nombre)})
                except Exception as ex:
                    pass

            if action == 'cargararchivoimagen':
                try:
                    data['title'] = u'Evidencias de requisitos de maestría'
                    data['id'] = request.GET['id']
                    data['idevidencia'] = request.GET['idevidencia']
                    requisito = RequisitosMaestria.objects.get(pk=int(request.GET['idevidencia']), status=True)
                    form = RequisitosMaestriaImgForm()
                    data['form'] = form
                    template = get_template("adm_admision/add_requisitomaestriaimagen.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content,
                                         "nombre": "SUBIR DOCUMENTO DE REQUISITO " + str(requisito.requisito.nombre)})
                except Exception as ex:
                    pass

            if action == 'addcohortes':
                try:
                    data['title'] = u'Adicionar Cohorte'
                    data['maestriaadmision'] = MaestriasAdmision.objects.get(
                        pk=int(encrypt(request.GET['idperiodomaestria'])))
                    form = CohorteMaestriaForm(
                        initial={'periodoacademico': periodo, 'modalidad': Modalidad.objects.get(pk=1)})
                    form.fields['coordinador'].queryset = Persona.objects.none()
                    data['form'] = form
                    template = get_template("adm_admision/addprogramamaestria.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'editarcohorte':
                try:
                    data['title'] = u'Editar Cohorte'
                    if persona.profesor():
                        data['personasesion'] = persona
                    else:
                        data['personasesion'] = persona
                    data['cohorte'] = cohorte = CohorteMaestria.objects.get(pk=int(encrypt(request.GET['id'])))
                    form = CohorteMaestriaForm(initial={
                        'periodoacademico': cohorte.periodoacademico,
                        'coordinador': cohorte.coordinador,
                        'descripcion': cohorte.descripcion,
                        'modalidad': cohorte.modalidad,
                        'alias': cohorte.alias,
                        'numerochorte': cohorte.numerochorte,
                        'cupodisponible': cohorte.cupodisponible,
                        # 'cuposlibres': cohorte.cuposlibres,
                        'cantidadgruposexamen': cohorte.cantidadgruposexamen,
                        'fechainiciocohorte': cohorte.fechainiciocohorte,
                        'fechafincohorte': cohorte.fechafincohorte,
                        'fechainicioinsp': cohorte.fechainicioinsp,
                        'fechafininsp': cohorte.fechafininsp,
                        'finiciorequisitos': cohorte.fechainiciorequisito,
                        'ffinrequisitos': cohorte.fechafinrequisito,
                        'fechafinrequisitobeca': cohorte.fechafinrequisitobeca,
                        'fechainicioexamen': cohorte.fechainicioexamen,
                        'fechafinexamen': cohorte.fechafinexamen,
                        'notaminimaexa': cohorte.notaminimaexa,
                        'notaminimatest': cohorte.notaminimatest,
                        'numerocuota': cohorte.numerocuota,
                        'valorcuota': cohorte.valorcuota,
                        'cantidadgruposentrevista': cohorte.cantidadgruposentrevista,
                        'minutosrango': cohorte.minutosrango,
                        'totaladmitidoscohorte': cohorte.totaladmitidoscohorte,
                        'activo': cohorte.activo,
                        'tienecostotramite': cohorte.tienecostotramite,
                        'valortramite': cohorte.valortramite,
                        'tipo': cohorte.tipo})
                    if cohorte.coordinador:
                        form.fields['coordinador'].queryset = Persona.objects.filter(pk=cohorte.coordinador.pk)
                    else:
                        form.fields['coordinador'].queryset = Persona.objects.none()
                    data['form2'] = form
                    template = get_template("adm_admision/editarprogramamaestria.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'addpreguntamaestria':
                try:
                    data['title'] = u'Adicionar Pregunta de Maestría'
                    data['maestria'] = CohorteMaestria.objects.get(pk=int(encrypt(request.GET['idmaestria'])))
                    form = AdmiPreguntasMaestriaForm()
                    data['form'] = form
                    return render(request, 'adm_admision/addpreguntamaestria.html', data)
                except Exception as ex:
                    pass

            if action == 'reportpdfpreinscritoscarrera':
                try:
                    data = {}
                    fecha = datetime.now()
                    format = fecha.strftime("%x")
                    data['fechaactual'] = u"%s" % format
                    persona = request.session['persona']
                    idc = MaestriasAdmision.objects.values_list('carrera_id', flat=True).filter(
                        cohortemaestria__coordinador=persona.id, status=True).distinct()
                    preinscritos = PreInscripcion.objects.filter(status=True).order_by('persona__apellido1',
                                                                                       'persona__apellido2')
                    if idc:
                        preinscritos = preinscritos.filter(carrera__in=idc)
                        if int(request.GET['id']) > 0:
                            data['preinscritos'] = preinscritos1 = preinscritos.filter(
                                carrera_id=int(request.GET['id']))
                    else:
                        data['preinscritos'] = preinscritos1 = preinscritos.filter(carrera_id=int(request.GET['id']))
                        if len(preinscritos1) == 0:
                            data['preinscritos'] = preinscritos1 = preinscritos.filter(status=True)
                            data['car'] = 0

                    return conviert_html_to_pdf(
                        'adm_admision/reportepdfpreinscritoscarrera.html',
                        {
                            'pagesize': 'A4',
                            'data': data,
                            'carrera': preinscritos1[0],
                        }
                    )
                except Exception as ex:
                    pass

            if action == 'editarmaestriaadmision':
                try:
                    data['maestriaadmision'] = maestriaadmision = MaestriasAdmision.objects.get(
                        pk=int(encrypt(request.GET['id'])))
                    form = AdmiPeriodoForm(initial={'descripcion': maestriaadmision.descripcion,
                                                    'carrera': maestriaadmision.carrera})
                    data['id'] = int(encrypt(request.GET['id']))
                    if maestriaadmision.en_uso():
                        form.editar()
                    data['form2'] = form
                    template = get_template('adm_admision/editarperiodoadmision.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'copiarrequisitos':
                try:
                    data['maestriaadmision'] = maestriaadmision = MaestriasAdmision.objects.get(
                        pk=int(encrypt(request.GET['id'])))
                    form = CopiarRequisitosForm()
                    form.fields['programa'].queryset = MaestriasAdmision.objects.filter(status=True, requisitosmaestria__isnull=False).exclude(pk=int(encrypt(request.GET['id']))).distinct()
                    data['id'] = int(encrypt(request.GET['id']))
                    data['form2'] = form
                    template = get_template('adm_admision/modal/copiarrequisitos.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'editarrequisitomaestria':
                try:
                    data['title'] = u'Editar requisito de maestría'
                    data['requisito'] = requisito = RequisitosMaestria.objects.get(
                        pk=int(encrypt(request.GET['idrequisitos'])))
                    form = AdmiRequisitosMaestriaForm(
                        initial={'descripcion': requisito.descripcion, 'observacion': requisito.observacion,
                                 'activo': requisito.activo, 'archivo': requisito.archivo,
                                 'requerido': requisito.requerido})
                    data['form'] = form
                    return render(request, "adm_admision/editarrequisitomaestria.html", data)
                except Exception as ex:
                    pass

            if action == 'addpagoexamen':
                try:
                    data['title'] = u'Pago de Examen'
                    data['cohorte'] = cohorte = CohorteMaestria.objects.get(pk=int(encrypt(request.GET['idcohorte'])))
                    data['inscripcioncohorte'] = InscripcionCohorte.objects.get(pk=int(encrypt(request.GET['id'])))
                    form = AdmiPagoExamenForm(initial={'valor': cohorte.valorexamen})
                    data['form'] = form
                    return render(request, "adm_admision/addpagoexamen.html", data)
                except Exception as ex:
                    pass

            if action == 'listadoaprobadorequisitos':
                try:
                    lista = []
                    idgruposex = IntegranteGrupoExamenMsc.objects.filter(status=True, inscripcion__cohortes__id=int(encrypt(request.GET['idcohorte']))).values_list('inscripcion__id', flat=True)
                    listadoaprobados = InscripcionCohorte.objects.filter(
                        cohortes_id=int(encrypt(request.GET['idcohorte'])),
                        estado_aprobador=2, status=True).exclude(id__in=idgruposex).order_by('inscripcionaspirante__persona__apellido1',
                                                                        'inscripcionaspirante__persona__apellido2')
                    for lis in listadoaprobados:
                        lista.append([lis.id,
                                      str(lis.inscripcionaspirante.persona.apellido1 + ' ' + lis.inscripcionaspirante.persona.apellido2 + ' ' + lis.inscripcionaspirante.persona.nombres)])
                    data = {"results": "ok", 'listadoaprobados': lista}
                    return JsonResponse(data)
                except Exception as ex:
                    pass

            if action == 'verlistadorequisitos':
                try:
                    lista = []
                    cohorte = CohorteMaestria.objects.get(pk=int(encrypt(request.GET['idcohorte'])), status=True)
                    if request.GET['clasereq'] == '1':
                        requisitosexcluir = ClaseRequisito.objects.values_list('requisito__id').filter(clasificacion=3,
                                                                                                       status=True)
                        requisitos = ClaseRequisito.objects.values_list('requisito__id').filter(clasificacion=1,
                                                                                                status=True).exclude(
                            requisito_id__in=requisitosexcluir)
                    else:
                        requisitosexcluir = ClaseRequisito.objects.values_list('requisito__id').filter(clasificacion=1,
                                                                                                       status=True)
                        requisitos = ClaseRequisito.objects.values_list('requisito__id').filter(clasificacion=3,
                                                                                                status=True).exclude(
                            requisito_id__in=requisitosexcluir)
                    requisitoscohorte = cohorte.requisitosmaestria_set.values_list('requisito_id').filter(status=True)
                    listadorequisitos = Requisito.objects.filter(pk__in=requisitos, activo=True, status=True).exclude(
                        pk__in=requisitoscohorte).order_by('nombre')
                    for lis in listadorequisitos:
                        lista.append([lis.id, lis.nombre, lis.observacion])
                    data = {"results": "ok", 'listadorequisitos': lista}
                    return JsonResponse(data)
                except Exception as ex:
                    pass

            if action == 'verlistadorequisitosmaestria':
                try:
                    data['eMaestria'] = MaestriasAdmision.objects.get(status=True, pk=int(request.GET['id']))
                    requisitosmaestria = RequisitosMaestria.objects.filter(status=True, maestria__id=int(request.GET['id'])).values_list('requisito__id', flat=True)
                    data['listadorequisitos'] = Requisito.objects.filter(status=True, claserequisito__clasificacion__id=4, activo=True).exclude(id__in=requisitosmaestria).order_by('nombre')
                    template = get_template("adm_admision/modal/listadorequisitosho.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'listadoaprobadoexamen':
                try:
                    lista = []
                    idgrupoent = IntegranteGrupoEntrevitaMsc.objects.filter(status=True, inscripcion__cohortes__id=int(encrypt(request.GET['idcohorte']))).values_list('inscripcion__id', flat=True)
                    listadoaprobados = InscripcionCohorte.objects.filter(
                        cohortes_id=int(encrypt(request.GET['idcohorte'])), status=True,
                        integrantegrupoexamenmsc__estado=2).exclude(id__in=idgrupoent).order_by(
                        'inscripcionaspirante__persona__apellido1', 'inscripcionaspirante__persona__apellido2')
                    for lis in listadoaprobados:
                        lista.append([lis.id,
                                      str(lis.inscripcionaspirante.persona.apellido1 + ' ' + lis.inscripcionaspirante.persona.apellido2 + ' ' + lis.inscripcionaspirante.persona.nombres)])
                    data = {"results": "ok", 'listadoaprobados': lista}
                    return JsonResponse(data)
                except Exception as ex:
                    pass

            if action == 'addmatricula':
                try:
                    data['title'] = u'PAGO DE MATRÍCULA'
                    data['cohorte'] = cohorte = CohorteMaestria.objects.get(pk=int(encrypt(request.GET['idcohorte'])))
                    data['inscripcioncohorte'] = InscripcionCohorte.objects.get(pk=int(encrypt(request.GET['id'])))
                    form = AdmiPagoMatriculaForm(initial={'valor': cohorte.valormatricula})
                    data['form'] = form
                    return render(request, "adm_admision/addmatricua.html", data)
                except Exception as ex:
                    pass

            if action == 'editarpreguntamaestria':
                try:
                    data['title'] = u'Editar pregunta de maestría'
                    data['pregunta'] = requisito = PreguntaMaestria.objects.get(
                        pk=int(encrypt(request.GET['idpregunta'])))
                    form = AdmiPreguntasMaestriaForm(
                        initial={'descripcion': requisito.descripcion, 'activo': requisito.activo})
                    data['form'] = form
                    return render(request, "adm_admision/editarpreguntamaestria.html", data)
                except Exception as ex:
                    pass

            if action == 'listadocohortes':
                try:
                    url_vars,estado,filtro = f'?action=listadocohortes&id={request.GET["id"]}', int(request.GET.get('estado',1)),Q(status=True)
                    if estado>0:
                        url_vars +='&estado='+str(estado)
                        estado_ve = True if estado == 1 else False
                        filtro = filtro & Q(procesoabierto=estado_ve)
                    data['estado'] = estado
                    data['url_vars'] = url_vars
                    data['title'] = u'Listado de Cohortes'
                    data['maestriaadmision'] = maestriaadmision = MaestriasAdmision.objects.get(
                        pk=int(encrypt(request.GET['id'])))
                    if puede_ver_todoadmision(request, 'posgrado.change_requisitosmaestria'):
                        data['listadocohortes'] = CohorteMaestria.objects.filter(maestriaadmision=maestriaadmision,
                                                                                 status=True).filter(filtro).order_by('id')
                    else:
                        if puede_ver_todoadmision(request, 'sga.puede_ver_todo_admisionposgrado'):
                            data['listadocohortes'] = CohorteMaestria.objects.filter(maestriaadmision=maestriaadmision,
                                                                                     status=True).filter(filtro).order_by('id')
                        else:
                            data['listadocohortes'] = CohorteMaestria.objects.filter(coordinador=persona,
                                                                                     maestriaadmision=maestriaadmision,
                                                                                     status=True).filter(filtro).order_by('id')
                    return render(request, "adm_admision/listamaestrias.html", data)
                except Exception as ex:
                    pass

            if action == 'configurarobsrubro':
                try:
                    data['cohorte'] = CohorteMaestria.objects.get(pk=int(encrypt(request.GET['id'])))
                    template = get_template("adm_admision/configurarobsrubro.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'planificarentrevista':
                try:
                    data['title'] = u'Planificar entrevista'
                    data['cohorte'] = cohorte = CohorteMaestria.objects.get(pk=int(encrypt(request.GET['id'])))
                    data['grupoentrevistas'] = cohorte.grupoentrevistamsc_set.filter(status=True).order_by('id')
                    return render(request, "adm_admision/gruposentrevista.html", data)
                except Exception as ex:
                    pass
            if action == 'matrizlistaexamen':
                try:
                    data['title'] = u'Subir listado'
                    data['cohorte'] = CohorteMaestria.objects.get(status=True,
                                                                  pk=int(encrypt(request.GET['idcohorte'])))
                    form = MatrizInscritosEntrevistaForm()
                    data['form'] = form
                    return render(request, "adm_admision/matrizlistaexamen.html", data)
                except Exception as ex:
                    pass

            if action == 'matrizlistanotaentrevista':
                try:
                    data[
                        'title'] = u'Subir listado notas entrevista, formato cedula(texto)|apellido(texto)|nombre(texto)|email(texto)|nota(numero)'
                    data['cohorte'] = CohorteMaestria.objects.get(status=True,
                                                                  pk=int(encrypt(request.GET['idcohorte'])))
                    form = MatrizInscritosEntrevistaForm()
                    data['form'] = form
                    return render(request, "adm_admision/matrizlistanotaentrevista.html", data)
                except Exception as ex:
                    pass

            if action == 'addgrupoentrevista':
                try:
                    data['title'] = u'Adicionar grupo entrevista'
                    data['cohorte'] = cohorte = CohorteMaestria.objects.get(status=True,
                                                                            pk=int(encrypt(request.GET['id'])))
                    data['form'] = GrupoEntrevistaForm()
                    return render(request, "adm_admision/addgrupoentrevista.html", data)
                except Exception as ex:
                    pass

            if action == 'editgrupoentrevista':
                try:
                    data['title'] = u'Editar grupo de Entrevista'
                    data['cohorte'] = cohorte = CohorteMaestria.objects.get(status=True,
                                                                            pk=int(encrypt(request.GET['idc'])))
                    data['grupo'] = grupo = GrupoEntrevistaMsc.objects.get(status=True,
                                                                           pk=int(encrypt(request.GET['id'])),
                                                                           cohortes=cohorte)
                    listaparticipantes = IntegranteGrupoEntrevitaMsc.objects.values_list("inscripcion_id",
                                                                                         flat=False).filter(status=True,
                                                                                                            grupoentrevista__cohortes=cohorte,
                                                                                                            inscripcion_status=True,
                                                                                                            inscripcion__activo=True,
                                                                                                            inscripcion__aproboproceso=False).exclude(
                        grupoentrevista=grupo)
                    listaentrevistados = RespuestaEntrevitaMsc.objects.values_list("integrante__inscripcion_id",
                                                                                   flat=False).filter(
                        integrante__inscripcion__cohortes=cohorte).distinct()
                    # data['participantes'] = IntegranteGrupoExamenMsc.objects.filter(status=True, inscripcion__activo=True, inscripcion__aproboproceso=False, grupoexamen__cohorte=cohorte, inscripcion__integrantegrupoexamenmsc__estado=2).exclude(inscripcion__in=listaparticipantes)
                    data['participantes'] = IntegranteGrupoExamenMsc.objects.filter(status=True,
                                                                                    inscripcion__activo=True,
                                                                                    grupoexamen__cohorte=cohorte,
                                                                                    inscripcion__integrantegrupoexamenmsc__estado=2).exclude(
                        inscripcion__in=listaparticipantes).exclude(inscripcion__in=listaentrevistados)
                    data['integrantesincritos'] = grupo.integrantegrupoentrevitamsc_set.filter(status=True)
                    # form = GrupoEntrevistaForm(initial={'administrativo':grupo.administrativo.id,
                    if grupo.administrativo:
                        form = GrupoEntrevistaForm(initial={'administrativo': grupo.administrativo.id,
                                                            'lugar': grupo.lugar,
                                                            'fecha': grupo.fecha,
                                                            'urlzoom': grupo.urlzoom,
                                                            'horainicio': grupo.horainicio,
                                                            'visible': grupo.visible,
                                                            'observacion': grupo.observacion})
                        if grupo.administrativo:
                            form.fields['administrativo'].widget.attrs['descripcion'] = grupo.administrativo
                            form.fields['administrativo'].widget.attrs['value'] = grupo.administrativo.id
                        else:
                            form.fields['administrativo'].widget.attrs['descripcion'] = '----------------'
                            form.fields['administrativo'].widget.attrs['value'] = 0
                    # form.editar()
                    else:
                        data['tiene_profesor'] = False
                        form = GrupoEntrevistaForm()
                    data['form'] = form
                    return render(request, "adm_admision/editgrupoentrevista.html", data)
                except Exception as ex:
                    pass

            if action == 'delintegranteentrevista':
                try:
                    data['title'] = u'Eliminar Integrante'
                    data['t'] = int(request.GET['t'])
                    data['integrante'] = IntegranteGrupoEntrevitaMsc.objects.get(pk=int(encrypt(request.GET['idi'])),
                                                                                 grupoentrevista_id=int(
                                                                                     encrypt(request.GET['idg'])))
                    return render(request, "adm_admision/delintegranteentrevista.html", data)
                except Exception as ex:
                    pass

            if action == 'delgrupoentrevista':
                try:
                    data['title'] = u'Eliminar grupo de entrevista'
                    data['grupo'] = grupo = GrupoEntrevistaMsc.objects.get(pk=int(encrypt(request.GET['id'])),
                                                                           cohortes_id=int(encrypt(request.GET['idc'])))
                    data['cohorte'] = grupo.cohortes
                    return render(request, "adm_admision/delgrupoentrevista.html", data)
                except Exception as ex:
                    pass

            if action == 'listaintegranteentrevista':
                try:
                    data['title'] = u'Lista de Integrantes'
                    data['grupo'] = grupo = GrupoEntrevistaMsc.objects.get(pk=int(encrypt(request.GET['id'])),
                                                                           cohortes_id=int(encrypt(request.GET['idc'])),
                                                                           status=True)
                    data['integrantes'] = grupo.integrantegrupoentrevitamsc_set.filter(inscripcion__status=True,
                                                                                       status=True).order_by(
                        'inscripcion__inscripcionaspirante__persona__apellido1',
                        'inscripcion__inscripcionaspirante__persona__apellido2')
                    return render(request, "adm_admision/entrevistados.html", data)
                except Exception as ex:
                    pass

            if action == 'listaintegranteexamennotas':
                try:
                    data['title'] = u'Lista de Integrantes Examen'
                    data['cohorte'] = cohorte = CohorteMaestria.objects.get(pk=int(encrypt(request.GET['idcohorte'])),
                                                                            status=True)
                    data['integrantes'] = IntegranteGrupoExamenMsc.objects.filter(grupoexamen__cohorte=cohorte,
                                                                                  status=True,
                                                                                  inscripcion__status=True).order_by(
                        'inscripcion__inscripcionaspirante__persona__apellido1',
                        'inscripcion__inscripcionaspirante__persona__apellido2')
                    return render(request, "adm_admision/listaintegranteexamennotas.html", data)
                except Exception as ex:
                    pass

            if action == 'listaintegranteentrevistanotas':
                try:
                    data['title'] = u'Lista de Integrantes'
                    data['cohorte'] = cohorte = CohorteMaestria.objects.get(pk=int(encrypt(request.GET['idcohorte'])),
                                                                            status=True)
                    data['listadocohortes'] = CohorteMaestria.objects.filter(maestriaadmision=cohorte.maestriaadmision,
                                                                             status=True).exclude(pk=cohorte.id)
                    data['integrantes'] = IntegranteGrupoEntrevitaMsc.objects.filter(grupoentrevista__cohortes=cohorte,
                                                                                     status=True,
                                                                                     inscripcion__status=True).order_by(
                        '-notafinal', 'inscripcion__inscripcionaspirante__persona__apellido1',
                        'inscripcion__inscripcionaspirante__persona__apellido2')
                    return render(request, "adm_admision/listaintegranteentrevistanotas.html", data)
                except Exception as ex:
                    pass

            if action == 'notasentrevista':
                try:
                    data['title'] = u'Lista de Integrantes'
                    data['cohorte'] = cohorte = CohorteMaestria.objects.get(pk=int(encrypt(request.GET['idc'])),
                                                                            status=True)
                    data['grupoentrevista'] = grupoentrevista = GrupoEntrevistaMsc.objects.get(
                        pk=int(encrypt(request.GET['id'])), status=True)
                    data['listadocohortes'] = CohorteMaestria.objects.filter(maestriaadmision=cohorte.maestriaadmision,
                                                                             status=True).exclude(pk=cohorte.id)
                    search = None
                    ids = None
                    if 's' in request.GET:
                        search = request.GET['s']
                        ss = search.split(' ')
                        if len(ss) == 1:
                            integrantes = IntegranteGrupoEntrevitaMsc.objects.filter(
                                Q(inscripcion__inscripcionaspirante__persona__nombres__icontains=search) |
                                Q(inscripcion__inscripcionaspirante__persona__apellido1__icontains=search) |
                                Q(inscripcion__inscripcionaspirante__persona__apellido2__icontains=search) |
                                Q(inscripcion__inscripcionaspirante__persona__cedula__icontains=search) |
                                Q(inscripcion__inscripcionaspirante__persona__pasaporte__icontains=search) |
                                Q(inscripcion__inscripcionaspirante__persona__usuario__username__icontains=search),
                                grupoentrevista=grupoentrevista, inscripcion__status=True,
                                grupoentrevista__cohortes=cohorte, status=True).distinct()
                        else:
                            integrantes = IntegranteGrupoEntrevitaMsc.objects.filter(
                                Q(inscripcion__inscripcionaspirante__persona__apellido1__icontains=ss[0]) &
                                Q(inscripcion__inscripcionaspirante__persona__apellido2__icontains=ss[1]),
                                grupoentrevista=grupoentrevista, inscripcion__status=True,
                                grupoentrevista__cohortes=cohorte, status=True).distinct()
                    else:
                        integrantes = IntegranteGrupoEntrevitaMsc.objects.filter(grupoentrevista=grupoentrevista,
                                                                                 inscripcion__status=True,
                                                                                 grupoentrevista__cohortes=cohorte,
                                                                                 status=True).order_by('-notafinal',
                                                                                                       'inscripcion__inscripcionaspirante__persona__apellido1',
                                                                                                       'inscripcion__inscripcionaspirante__persona__apellido2')
                    numerofilas = 25
                    paging = MiPaginador(integrantes, numerofilas)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                            if p == 1:
                                numerofilasguiente = numerofilas
                            else:
                                numerofilasguiente = numerofilas * (p - 1)
                        else:
                            p = paginasesion
                            if p == 1:
                                numerofilasguiente = numerofilas
                            else:
                                numerofilasguiente = numerofilas * (p - 1)
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['page'] = page
                    data['numerofilasguiente'] = numerofilasguiente
                    data['numeropagina'] = p
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['integrantes'] = page.object_list
                    data['search'] = search if search else ""
                    return render(request, "adm_admision/notasentrevista.html", data)
                except Exception as ex:
                    pass

            if action == 'listadorequisitos':
                try:
                    data['title'] = u'Listado de Requisitos'
                    hoy = datetime.now().date()
                    data['puede_adicionar'] = False
                    data['cohorte'] = cohorte = CohorteMaestria.objects.get(pk=int(encrypt(request.GET['id'])))
                    if cohorte.fechainiciorequisito > hoy:
                        data['puede_adicionar'] = True
                    requisitosadmision = ClaseRequisito.objects.values_list('requisito__id').filter(clasificacion=1,
                                                                                                    status=True)
                    requisitoscomercializacion = ClaseRequisito.objects.values_list('requisito__id').filter(
                        clasificacion=3, status=True)
                    data['requisitosadmision'] = cohorte.requisitosmaestria_set.filter(requisito__in=requisitosadmision,
                                                                                       status=True).order_by('id')
                    data['requisitoscomercializacion'] = cohorte.requisitosmaestria_set.filter(
                        requisito__in=requisitoscomercializacion, status=True).exclude(
                        requisito__in=requisitosadmision).order_by('id')
                    return render(request, "adm_admision/listarequisitos.html", data)
                except Exception as ex:
                    pass

            if action == 'listadoadmitidosconproceso':
                try:
                    data['title'] = u'Listado de Admitidos'
                    hoy = datetime.now().date()
                    search = None
                    ids = None
                    data['cohorte'] = cohorte = CohorteMaestria.objects.get(pk=int(encrypt(request.GET['id'])))
                    totalesconrubro = 0
                    listadointegrantes = IntegranteGrupoEntrevitaMsc.objects.filter(estado_emailadmitido=2,
                                                                                    cohorteadmitidasinproceso__isnull=True,
                                                                                    grupoentrevista__cohortes=cohorte,
                                                                                    status=True,
                                                                                    inscripcion__status=True).select_related(
                        'inscripcion__inscripcionaspirante__persona')
                    data['totalsingenerar'] = listadointegrantes.filter(inscripcion__tipocobro=1).count()
                    totalesconrubro = Rubro.objects.filter(
                        inscripcion_id__in=listadointegrantes.values_list('inscripcion_id'),
                        admisionposgradotipo__in=[2, 3], status=True).count()
                    data['totalesconrubro'] = totalesconrubro
                    data['listadointegrantes'] = listadointegrantes.count()
                    if 's' in request.GET:
                        search = request.GET['s'].strip()
                        ss = search.split(' ')
                        if len(ss) == 1:
                            listadoadmitidossinproceso = IntegranteGrupoEntrevitaMsc.objects.select_related().filter(
                                Q(inscripcion__inscripcionaspirante__persona__nombres__icontains=search) |
                                Q(inscripcion__inscripcionaspirante__persona__apellido1__icontains=search) |
                                Q(inscripcion__inscripcionaspirante__persona__apellido2__icontains=search) |
                                Q(inscripcion__inscripcionaspirante__persona__cedula__icontains=search) |
                                Q(inscripcion__inscripcionaspirante__persona__pasaporte__icontains=search) |
                                Q(inscripcion__inscripcionaspirante__persona__usuario__username__icontains=search),
                                estado_emailadmitido=2, cohorteadmitidasinproceso__isnull=True,
                                grupoentrevista__cohortes=cohorte, status=True,
                                inscripcion__status=True).select_related(
                                'inscripcion__inscripcionaspirante__persona').distinct()
                        else:
                            listadoadmitidossinproceso = IntegranteGrupoEntrevitaMsc.objects.select_related().filter(
                                Q(inscripcion__inscripcionaspirante__persona__apellido1__icontains=ss[0]) &
                                Q(inscripcion__inscripcionaspirante__persona__apellido2__icontains=ss[1]),
                                estado_emailadmitido=2, cohorteadmitidasinproceso__isnull=True,
                                grupoentrevista__cohortes=cohorte, status=True,
                                inscripcion__status=True).select_related(
                                'inscripcion__inscripcionaspirante__persona').distinct()
                    else:
                        listadoadmitidossinproceso = IntegranteGrupoEntrevitaMsc.objects.filter(estado_emailadmitido=2,
                                                                                                cohorteadmitidasinproceso__isnull=True,
                                                                                                grupoentrevista__cohortes=cohorte,
                                                                                                status=True,
                                                                                                inscripcion__status=True).order_by(
                            'inscripcion__inscripcionaspirante__persona__apellido1',
                            'inscripcion__inscripcionaspirante__persona__apellido2').select_related(
                            'inscripcion__inscripcionaspirante__persona')
                    numerofilas = 20
                    paging = MiPaginador(listadoadmitidossinproceso, numerofilas)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                            if p == 1:
                                numerofilasguiente = numerofilas
                            else:
                                numerofilasguiente = numerofilas * (p - 1)
                        else:
                            p = paginasesion
                            if p == 1:
                                numerofilasguiente = numerofilas
                            else:
                                numerofilasguiente = numerofilas * (p - 1)
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['numerofilasguiente'] = numerofilasguiente
                    data['numeropagina'] = p
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['listadoadmitidossinproceso'] = page.object_list
                    return render(request, "adm_admision/listadoadmitidossinproceso.html", data)
                except Exception as ex:
                    pass

            if action == 'listadoadmitidoscohorte':
                try:
                    data['title'] = u'Listado de Admitidos'
                    hoy = datetime.now().date()
                    search = None
                    ids = None
                    data['cohorte'] = cohorte = CohorteMaestria.objects.get(pk=int(encrypt(request.GET['id'])))
                    if cohorte.maestriaadmision.carrera.malla().tiene_itinerario_malla_especialidad():
                        data['menciones'] = ItinerarioMallaEspecilidad.objects.filter(status=True,
                                                                                      malla__id=cohorte.maestriaadmision.carrera.malla().id)
                    totalesconrubro = 0
                    if cohorte.tipo == 2:
                        listadointegrantes = InscripcionCohorte.objects.filter(cohortes=cohorte, cohortes__tipo=2,
                                                                               estado_aprobador=2,
                                                                               integrantegrupoexamenmsc__estado=2,
                                                                               status=True)
                    if cohorte.tipo == 3:
                        listadointegrantes = InscripcionCohorte.objects.filter(cohortes=cohorte, cohortes__tipo=3,
                                                                               estado_aprobador=2, status=True)
                    listadosingenerar = 0
                    for integ in listadointegrantes:
                        if not integ.genero_rubro_programa():
                            listadosingenerar += 1
                    data['totalsingenerar'] = listadosingenerar
                    # data['totalsingenerar'] = listadointegrantes.filter(tipocobro=1).count()
                    totalesconrubro = Rubro.objects.filter(inscripcion_id__in=listadointegrantes.values_list('id'),
                                                           admisionposgradotipo__in=[2, 3], cohortemaestria=cohorte,
                                                           status=True).count()
                    data['totalesconrubro'] = totalesconrubro
                    data['listadointegrantes'] = listadointegrantes.count()
                    if cohorte.tipo == 2:
                        if 's' in request.GET:
                            search = request.GET['s'].strip()
                            ss = search.split(' ')
                            if len(ss) == 1:
                                listadoadmitidossinproceso = InscripcionCohorte.objects.select_related().filter(
                                    Q(inscripcionaspirante__persona__nombres__icontains=search) |
                                    Q(inscripcionaspirante__persona__apellido1__icontains=search) |
                                    Q(inscripcionaspirante__persona__apellido2__icontains=search) |
                                    Q(inscripcionaspirante__persona__cedula__icontains=search) |
                                    Q(inscripcionaspirante__persona__pasaporte__icontains=search) |
                                    Q(inscripcionaspirante__persona__usuario__username__icontains=search),
                                    integrantegrupoexamenmsc__estado=2, cohortes=cohorte, estado_aprobador=2,
                                    status=True).distinct()
                            else:
                                listadoadmitidossinproceso = InscripcionCohorte.objects.select_related().filter(
                                    Q(inscripcionaspirante__persona__apellido1__icontains=ss[0]) &
                                    Q(inscripcionaspirante__persona__apellido2__icontains=ss[1]),
                                    integrantegrupoexamenmsc__estado=2, cohortes=cohorte, estado_aprobador=2,
                                    status=True).distinct()
                        else:
                            listadoadmitidossinproceso = InscripcionCohorte.objects.filter(cohortes=cohorte,
                                                                                           estado_aprobador=2,
                                                                                           status=True,
                                                                                           integrantegrupoexamenmsc__estado=2).order_by(
                                'inscripcionaspirante__persona__apellido1', 'inscripcionaspirante__persona__apellido2')

                    if cohorte.tipo == 3:
                        if 's' in request.GET:
                            search = request.GET['s'].strip()
                            ss = search.split(' ')
                            if len(ss) == 1:
                                listadoadmitidossinproceso = InscripcionCohorte.objects.select_related().filter(
                                    Q(inscripcionaspirante__persona__nombres__icontains=search) |
                                    Q(inscripcionaspirante__persona__apellido1__icontains=search) |
                                    Q(inscripcionaspirante__persona__apellido2__icontains=search) |
                                    Q(inscripcionaspirante__persona__cedula__icontains=search) |
                                    Q(inscripcionaspirante__persona__pasaporte__icontains=search) |
                                    Q(inscripcionaspirante__persona__usuario__username__icontains=search),
                                    cohortes=cohorte, estado_aprobador=2, status=True).distinct()
                            else:
                                listadoadmitidossinproceso = InscripcionCohorte.objects.select_related().filter(
                                    Q(inscripcionaspirante__persona__apellido1__icontains=ss[0]) &
                                    Q(inscripcionaspirante__persona__apellido2__icontains=ss[1]),
                                    cohortes=cohorte, estado_aprobador=2, status=True).distinct()
                        else:
                            listadoadmitidossinproceso = InscripcionCohorte.objects.filter(cohortes=cohorte,
                                                                                           estado_aprobador=2,
                                                                                           status=True).order_by(
                                'inscripcionaspirante__persona__apellido1', 'inscripcionaspirante__persona__apellido2')

                    listaconrubros = InscripcionCohorte.objects.filter(cohortes=cohorte, estado_aprobador=2,
                                                                       status=True, tipocobro__in=[2, 3]).order_by(
                        'inscripcionaspirante__persona__apellido1', 'inscripcionaspirante__persona__apellido2')
                    if not 's' in request.GET:
                        listadocompleto = listadoadmitidossinproceso | listaconrubros
                    else:
                        listadocompleto = listadoadmitidossinproceso

                    mencionselect = 0
                    if cohorte.maestriaadmision.carrera.malla().tiene_itinerario_malla_especialidad():
                        if 'mencion' in request.GET:
                            mencionselect = int(request.GET['mencion'])

                            if mencionselect > 0:
                                if mencionselect == 3:
                                    listadocompleto = listadocompleto.filter(itinerario=0)
                                else:
                                    listadocompleto = listadocompleto.filter(itinerario=mencionselect)
                            else:
                                listadocompleto = listadocompleto

                    numerofilas = 20
                    paging = MiPaginador(listadocompleto, numerofilas)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                            if p == 1:
                                numerofilasguiente = numerofilas
                            else:
                                numerofilasguiente = numerofilas * (p - 1)
                        else:
                            p = paginasesion
                            if p == 1:
                                numerofilasguiente = numerofilas
                            else:
                                numerofilasguiente = numerofilas * (p - 1)
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['numerofilasguiente'] = numerofilasguiente
                    data['numeropagina'] = p
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['listadoadmitidoscohorte'] = page.object_list
                    data['mencionselect'] = mencionselect
                    return render(request, "adm_admision/listadoadmitidoscohorte.html", data)
                except Exception as ex:
                    pass

            if action == 'listadoadmitidossinproceso':
                try:
                    data['title'] = u'Listado de Requisitos'
                    hoy = datetime.now().date()
                    data['cohorte'] = cohorte = CohorteMaestria.objects.get(pk=int(encrypt(request.GET['id'])))
                    data['listadoadmitidossinproceso'] = cohorte.integrantegrupoentrevitamsc_set.filter(status=True,
                                                                                                        inscripcion__status=True).order_by(
                        'inscripcion__inscripcionaspirante__persona__apellido1',
                        'inscripcion__inscripcionaspirante__persona__apellido2')
                    return render(request, "adm_admision/listadoadmitidossinproceso.html", data)
                except Exception as ex:
                    pass

            if action == 'listadoaprobadosentrevista':
                try:
                    data['title'] = u'Listado de aprobados entrevista'
                    data['cohorte'] = cohorte = CohorteMaestria.objects.get(pk=int(encrypt(request.GET['id'])))
                    data['integrantes'] = IntegranteGrupoEntrevitaMsc.objects.filter(grupoentrevista__cohortes=cohorte,
                                                                                     estado=2, status=True,
                                                                                     inscripcion__status=True)
                    return render(request, "adm_admision/aprobadosentrevista.html", data)
                except Exception as ex:
                    pass

            if action == 'listacuotas':
                try:
                    data['title'] = u'Listado de coutas'
                    data['inscripcioncohorte'] = inscripcioncohorte = InscripcionCohorte.objects.get(
                        pk=int(encrypt(request.GET['id'])))
                    data['totalcancelado'] = totalcancelado = \
                        CuotaPago.objects.filter(pago__inscripcioncohorte=inscripcioncohorte, status=True).aggregate(
                            valor=Sum('valor'))['valor']
                    data['totalfaltante'] = (
                                                    inscripcioncohorte.cohortes.costo_maestria() + inscripcioncohorte.cohortes.valorexamen + inscripcioncohorte.cohortes.valormatricula) - totalcancelado
                    data['listacuotas'] = inscripcioncohorte.pago_set.filter(tipo_id=3, status=True).order_by(
                        'numerocuota')
                    return render(request, "adm_admision/listacuotas.html", data)
                except Exception as ex:
                    pass

            if action == 'addpagocuota':
                try:
                    data['title'] = u'Pago de Cuotas'
                    data['pagocuota'] = pago = Pago.objects.get(pk=int(encrypt(request.GET['id'])))
                    form = MatriculaPagoCuota(initial={'valor': pago.inscripcioncohorte.cohortes.valorcuota})
                    data['form'] = form
                    return render(request, "adm_admision/addpagocuota.html", data)
                except Exception as ex:
                    pass

            if action == 'listadoaspirantes':
                try:
                    data['title'] = u'Listado de Inscritos'
                    data['persona'] = persona
                    fechadesde = ''
                    fechahasta = ''
                    inscripcioncohorte = None
                    data['cohorte'] = cohorte = CohorteMaestria.objects.get(pk=int(encrypt(request.GET['id'])))
                    if cohorte.maestriaadmision.carrera.malla().tiene_itinerario_malla_especialidad():
                        data['menciones'] = ItinerarioMallaEspecilidad.objects.filter(status=True, malla__id=cohorte.maestriaadmision.carrera.malla().id)

                    data['listadocohortes'] = cohorte.maestriaadmision.cohortemaestria_set.filter(status=True).exclude(
                        pk=cohorte.id).order_by('id')
                    data['id_cohorte'] = int(encrypt(request.GET['id']))
                    filter = Q(status=True)
                    if 'tipoestado' in request.GET:
                        data['tipoestado'] = int(request.GET['tipoestado'])
                        inscripcioncohorte = InscripcionCohorte.objects.filter(estado_aprobador=request.GET['tipoestado'],
                                                                       cohortes=cohorte, status=True)
                        aspirantes = inscripcioncohorte.select_related('inscripcionaspirante__persona').order_by( 'id',
                            'estado_emailevidencia', 'inscripcionaspirante__persona__apellido1',
                            'inscripcionaspirante__persona__apellido2', 'inscripcionaspirante__persona__nombres')
                        if int(request.GET['tipoestado']) == 1:
                            filter = Q(status=True, preaprobado=True)
                    else:
                        data['tipoestado'] = 1
                        inscripcioncohorte = InscripcionCohorte.objects.filter(estado_aprobador=1, cohortes=cohorte,
                                                                       status=True)
                        aspirantes = inscripcioncohorte.select_related('inscripcionaspirante__persona').order_by( 'id',
                            'inscripcionaspirante__persona__apellido1', 'inscripcionaspirante__persona__apellido2',
                            'inscripcionaspirante__persona__nombres')
                    data['totalrecaudado'] = CuotaPago.objects.filter(pago__inscripcioncohorte__cohortes=cohorte,
                                                                      pago__tipo_id=2, pago__status=True).aggregate(
                        valor=Sum('valor'))
                    # data['totalinscritos'] = aspirantes.values('id').count()

                    data['estados'] = ESTADO_REVISION
                    search = None
                    ids = None
                    mencionselect = 0
                    if 's' in request.GET:
                        search = request.GET['s']
                        ss = search.split(' ')
                        if len(ss) == 1:
                            # aspirantes = aspirantes.filter(
                            #     Q(inscripcionaspirante__persona__nombres__icontains=search) | Q(
                            #         inscripcionaspirante__persona__apellido1__icontains=search) | Q(
                            #         inscripcionaspirante__persona__apellido2__icontains=search) | Q(
                            #         inscripcionaspirante__persona__cedula__icontains=search)).select_related('inscripcionaspirante__persona')
                            filter = filter & (Q(inscripcionaspirante__persona__nombres__icontains=search) | Q(
                                     inscripcionaspirante__persona__apellido1__icontains=search) | Q(
                                     inscripcionaspirante__persona__apellido2__icontains=search) | Q(
                                     inscripcionaspirante__persona__cedula__icontains=search))
                        else:
                            # aspirantes = aspirantes.filter((Q(inscripcionaspirante__persona__nombres__icontains=ss[0]) & Q(inscripcionaspirante__persona__nombres__icontains=ss[1])) | (Q(inscripcionaspirante__persona__apellido1__icontains=ss[0]) & Q(inscripcionaspirante__persona__apellido2__icontains=ss[1]))).select_related('inscripcionaspirante__persona')
                            aspirantes.select_related('inscripcionaspirante__persona')
                            filter = filter & ((Q(inscripcionaspirante__persona__nombres__icontains=ss[0]) & Q(inscripcionaspirante__persona__nombres__icontains=ss[1])) | (Q(inscripcionaspirante__persona__apellido1__icontains=ss[0]) & Q(inscripcionaspirante__persona__apellido2__icontains=ss[1])))
                    if 'fechadesde' in request.GET:
                        fechadesde = request.GET['fechadesde']
                        if fechadesde == '':
                            fechadesde = datetime.now().date().strftime("%d-%m-%Y")
                        fechahasta = request.GET['fechahasta']
                        aspirantes.select_related('inscripcionaspirante__persona')
                        # aspirantes = aspirantes.filter(fecha_creacion__date__range=(convertir_fecha(fechadesde), convertir_fecha(fechahasta))).select_related('inscripcionaspirante__persona')
                        filter = filter & Q(fecha_creacion__date__range=(convertir_fecha(fechadesde), convertir_fecha(fechahasta)))
                    if 'idi' in request.GET:
                        ids = int(encrypt(request.GET['idi']))
                        # aspirantes = aspirantes.filter(pk=int(encrypt(request.GET['idi'])))
                        filter = filter & Q(pk=int(encrypt(request.GET['idi'])))

                    if cohorte.maestriaadmision.carrera.malla().tiene_itinerario_malla_especialidad():
                        if 'mencion' in request.GET:
                            mencionselect = int(request.GET['mencion'])

                            if mencionselect > 0:
                                if mencionselect == 3:
                                    # aspirantes = aspirantes.filter(itinerario=0)
                                    filter = filter & Q(itinerario=0)
                                else:
                                    # aspirantes = aspirantes.filter(itinerario=mencionselect)
                                    filter = filter & Q(itinerario=mencionselect)
                            else:
                                aspirantes = aspirantes

                    if aspirantes:
                        aspirantes = aspirantes.filter(filter)
                    if 'cant_requisitos' in request.GET:
                        if request.GET['cant_requisitos']:
                            data['cantseleccionada'] = cant_requisitos = int(request.GET['cant_requisitos'])
                            ids = []
                            for inscrip in inscripcioncohorte.values_list('id', 'cohortes_id').distinct():
                                requisitos_cohorte = RequisitosMaestria.objects.filter(status=True, cohorte_id=inscrip[1]).values_list('id', flat=True).exclude(requisito__claserequisito__clasificacion=3)
                                evidencias = EvidenciaRequisitosAspirante.objects.filter(
                                    status=True, inscripcioncohorte_id=int(inscrip[0]), requisitos__id__in=requisitos_cohorte
                                ).values_list('id', flat=True)
                                if (cant_requisitos > 0 and evidencias.count() >=cant_requisitos ):
                                    # num = 0
                                    # for evidencia in evidencias:
                                    #     num += 1 if evidencia.detalleevidenciarequisitosaspirante_set.filter(Q(status=True) & ~Q(estado_aprobacion=3)) else 0
                                    #     if (num > cant_requisitos):
                                    #         break
                                    num = DetalleEvidenciaRequisitosAspirante.objects.values('id').filter(
                                        Q(evidencia__in=evidencias) & (
                                            Q(status=True) & ~Q(estado_aprobacion=3))).distinct().count()
                                    if (num == cant_requisitos):
                                        ids.append(int(inscrip[0]))

                                if (cant_requisitos == 0 and evidencias.count() == 0):
                                    ids.append(int(inscrip[0]))
                            aspirantes = aspirantes.filter(id__in=ids)


                    else:
                        data['requisitodefault'] = inscripcioncohorte.first().total_evidenciasgrupocohorteobligatorias() if inscripcioncohorte else None

                    #Ordenar por fecha
                    if aspirantes:
                        dict_fechas = []
                        for aspi in aspirantes:
                            d = {'id': aspi.id, 'dias': aspi.dias_transcurridos()}
                            dict_fechas.append(d)

                        newlist = sorted(dict_fechas, key=itemgetter('dias'), reverse=True)
                        lista_oficial = []
                        for n in newlist:
                            lista_oficial.append(n['id'])

                        orderi = Case(*[When(pk=pk, then=pos) for pos, pk in enumerate(lista_oficial)])
                        aspirantes = aspirantes.order_by(orderi)

                    paging = MiPaginador(aspirantes, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['page'] = page
                    data['fechadesde'] = fechadesde
                    data['fechahasta'] = fechahasta
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['listadoaspirantes'] = page.object_list
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['totalinscritos'] = aspirantes.values('id').count() if aspirantes else 0
                    data['lista_cant_requisitos'] = [x for x in range(cohorte.cant_requisitos() + 1)]
                    data['mencionselect'] = mencionselect
                    return render(request, "adm_admision/listadoaspirantes.html", data)
                except Exception as ex:
                    pass

            if action == 'listadopreguntas':
                try:
                    data['title'] = u'Listado de Preguntas para Programas Maestrías'
                    data['cohorte'] = cohorte = CohorteMaestria.objects.get(pk=int(encrypt(request.GET['id'])))
                    data['listadopreguntas'] = PreguntaMaestria.objects.filter(cohortes=cohorte, status=True).order_by(
                        'pregunta__tipopregunta')
                    return render(request, "adm_admision/listadopreguntas.html", data)
                except Exception as ex:
                    pass

            if action == 'informeevidencias':
                try:
                    data['title'] = u'Revisión de evidencias'
                    data['cohorte'] = cohorte = CohorteMaestria.objects.get(pk=int(request.GET['idcohorte']))
                    data['inscripcioncohorte'] = aspirante = \
                        cohorte.inscripcioncohorte_set.filter(inscripcionaspirante_id=request.GET['aspirante'],
                                                              status=True)[0]
                    data['requisitos'] = requisitos = cohorte.requisitosmaestria_set.filter(status=True).order_by('id')
                    data['evidencia'] = requisitos[0].detalle_requisitosmaestriacohorte(aspirante)
                    data['estados'] = ESTADO_REVISION
                    template = get_template("adm_admision/informe_evidencia.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            if action == 'informeevidencias1':
                try:
                    data['title'] = u'Revisión de evidencias'
                    requisito = RequisitosMaestria.objects.get(pk=request.GET['idrequisito'])
                    if not requisito.evidenciarequisitosaspirante_set.values("id").filter(
                            inscripcioncohorte_id=request.GET['idinscripcioncohote'], status=True).exists():
                        return JsonResponse({"result": "sin", "mensaje": u"NO EXISTE EVIDENCIA."})
                    data['requisitoinscrito'] = requisito.evidenciarequisitosaspirante_set.filter(
                        inscripcioncohorte_id=request.GET['idinscripcioncohote'], status=True).order_by('-id').first()
                    template = get_template("adm_admision/informe_evidencia.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            if action == 'informeevidenciascontrato':
                try:
                    data['title'] = u'Revisión de evidencia de contrato'
                    contrato = None
                    if 'idcontrato' in request.GET:
                        if request.GET['idcontrato']:
                            data['contrato'] = contrato = Contrato.objects.get(pk=request.GET['idcontrato'])
                    if not contrato:
                        return JsonResponse({"result": "sin", "mensaje": u"NO EXISTE EVIDENCIA."})
                    if not contrato.archivocontrato:
                        return JsonResponse({"result": "sin", "mensaje": u"NO EXISTE EVIDENCIA."})
                    template = get_template("adm_admision/informe_evidencia_contrato.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            if action == 'informeevidenciaspagare':
                try:
                    data['title'] = u'Revisión de evidencia de pagaré'
                    contrato = None
                    if 'idcontrato' in request.GET:
                        if request.GET['idcontrato']:
                            data['pagare'] = contrato = Contrato.objects.get(pk=request.GET['idcontrato'])
                    if not contrato:
                        return JsonResponse({"result": "sin", "mensaje": u"NO EXISTE EVIDENCIA."})
                    if not contrato.archivopagare:
                        return JsonResponse({"result": "sin", "mensaje": u"NO EXISTE EVIDENCIA."})
                    template = get_template("adm_admision/informe_evidencia_pagare.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            if action == 'tablaamortizacion':
                try:
                    data['title'] = u'Tabla de amortización'
                    data['tipoestado'] = request.GET['tipoestado']
                    data['cohorte'] = cohorte = CohorteMaestria.objects.get(status=True,
                                                                            pk=int(encrypt(request.GET['idcohorte'])))
                    data['aspirante'] = inscripcion = InscripcionCohorte.objects.get(status=True, pk=int(
                        encrypt(request.GET['aspirante'])))
                    # infofinancieramae = InfraestructuraEquipamientoInformacionPac.objects.filter(programapac__carrera=inscripcion.cohortes.maestriaadmision.carrera).last()
                    # if infofinancieramae.valorarancel:
                    #     data['infofinancieramae'] = infofinancieramae
                    data['tablaamortizacion'] = tablaamortizacion = TablaAmortizacion.objects.filter(status=True,
                                                                                                     contrato=inscripcion.contrato_set.last())
                    total = 0
                    for valor in tablaamortizacion:
                        total = total + valor.valor
                    data['total'] = total
                    return render(request, "adm_admision/tablaamortizacion.html", data)
                except Exception as ex:
                    pass

            if action == 'evidenciasinscritos':
                try:
                    data['title'] = u'Revisión de evidencias'
                    data['tipoestado'] = request.GET['tipoestado']
                    data['cant_requisitos'] = request.GET['cant_requisitos']
                    data['cohorte'] = cohorte = CohorteMaestria.objects.get(pk=int(encrypt(request.GET['idcohorte'])))
                    if cohorte.coordinador == persona:
                        data['revisaevidencia'] = True
                    else:
                        data['revisaevidencia'] = False
                    data['inscripcioncohorte'] = inscripcioncohorte = InscripcionCohorte.objects.get(
                        pk=int(encrypt(request.GET['aspirante'])), status=True)

                    if 'comercia' in request.GET:
                        requisitosexcluir = ClaseRequisito.objects.values_list('requisito__id').filter(clasificacion=1,
                                                                                                       status=True)
                        data['requisitos'] = cohorte.requisitosmaestria_set.filter(status=True).exclude(
                            requisito__in=requisitosexcluir).order_by('id')
                        data['comercia'] = True
                    else:
                        requisitosexcluir = ClaseRequisito.objects.values_list('requisito__id').filter(clasificacion=3,
                                                                                                       status=True)
                        if not inscripcioncohorte.grupo:
                            data['requisitos'] = cohorte.requisitosmaestria_set.filter(status=True).exclude(
                                requisito__in=requisitosexcluir).order_by('id')
                        else:
                            gruporequisitos = inscripcioncohorte.grupo.requisitosgrupocohorte_set.values_list(
                                'requisito_id', flat=True).filter(status=True)
                            data['requisitos'] = cohorte.requisitosmaestria_set.filter(requisito_id__in=gruporequisitos,
                                                                                       status=True).order_by('id')
                    return render(request, "adm_admision/evidenciasinscritos.html", data)
                except Exception as ex:
                    pass

            if action == 'historialrevision':
                try:
                    if 'id' in request.GET:
                        data['evidencia'] = evidencia = EvidenciaRequisitosAspirante.objects.get(pk=request.GET['id'],
                                                                                                 status=True)
                        data['historial'] = DetalleEvidenciaRequisitosAspirante.objects.filter(evidencia=evidencia,
                                                                                               status=True)
                    template = get_template("adm_admision/modal/historialrevisionrequisitos.html")
                    return JsonResponse({"result": 'ok', 'data': template.render(data)})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            if action == 'evidenciacontrato':
                try:
                    data['title'] = u'Revisión de contrato de pago'
                    data['tipoestado'] = request.GET['tipoestado']
                    data['cohorte'] = cohorte = CohorteMaestria.objects.get(pk=int(encrypt(request.GET['idcohorte'])))
                    if cohorte.coordinador == persona:
                        data['revisaevidencia'] = True
                    else:
                        data['revisaevidencia'] = False
                    data['inscripcioncohorte'] = inscripcioncohorte = InscripcionCohorte.objects.get(
                        pk=int(encrypt(request.GET['aspirante'])), status=True)
                    data['contrato'] = inscripcioncohorte.contrato_set.filter(status=True).last()
                    return render(request, "adm_admision/evidenciacontrato.html", data)
                except Exception as ex:
                    pass

            if action == 'adm_subirevidenciacontrato':
                try:
                    data['title'] = u'Registro de contrato de pago'
                    data['tipoestado'] = request.GET['tipoestado']
                    data['inscripcioncohorte'] = inscripcioncohorte = InscripcionCohorte.objects.get(
                        pk=int(encrypt(request.GET['idinscripcioncohorte'])))
                    permisocontrato = False
                    fechafincontrato = inscripcioncohorte.cohortes.fechafininsp + timedelta(days=30)
                    if fechafincontrato >= hoy:
                        permisocontrato = True
                    data['permisocontrato'] = permisocontrato
                    formapago = 1
                    if inscripcioncohorte.formapagopac:
                        formapago = inscripcioncohorte.formapagopac.id
                    if TipoFormaPagoPac.objects.filter(pk=formapago).exists():
                        data['fpago'] = fpago = TipoFormaPagoPac.objects.filter(pk=formapago).last()
                        data['contrato'] = contrato = Contrato.objects.filter(status=True,
                                                                              inscripcion=inscripcioncohorte,
                                                                              inscripcion__status=True).last()
                    return render(request, "adm_admision/adm_subircontratopago.html", data)
                except Exception as ex:
                    pass

            if action == 'cargarcontratopago':
                try:
                    data['title'] = u'Contrato de pago'
                    data['id'] = request.GET['id']
                    data['inscripcioncohorte'] = inscripcioncohorte = InscripcionCohorte.objects.get(
                        pk=request.GET['id'])
                    data['tipoestado'] = request.GET['tipoestado']
                    data['fpago'] = request.GET['fp']
                    form = ContratoPagoMaestriaForm()
                    data['form'] = form
                    template = get_template("alu_requisitosmaestria/adm_addcontratopago.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content, "nombre": "SUBIR CONTRATO DE PAGO"})
                except Exception as ex:
                    pass

            if action == 'listadorequisitosinscripcion':
                try:
                    data['title'] = u'Requisitos de programas de Maestría'
                    tienepagoexamen = False
                    ventanaactiva = 1
                    bloqueasubidapago = 1
                    hoy = datetime.now().date()
                    tienerequisitos = False
                    data['tipoestado'] = request.GET['tipoestado']
                    data['cant_requisitos'] = request.GET['cant_requisitos']
                    data['inscripcioncohorte'] = inscripcioncohorte = InscripcionCohorte.objects.get(
                        pk=int(encrypt(request.GET['idinscripcioncohorte'])))
                    inscripciondescuento = None
                    tienerequisitosbecas = False
                    data['tienerequisitosbecas'] = tienerequisitosbecas
                    data['inscripciondescuento'] = inscripciondescuento
                    if inscripcioncohorte.evidenciarequisitosaspirante_set.filter(status=True).exists():
                        tienerequisitos = True
                    data['tienerequisitos'] = tienerequisitos
                    if inscripcioncohorte.cohortes.tienecostoexamen:
                        if inscripcioncohorte.evidenciapagoexamen_set.filter(status=True):
                            tienepagoexamen = True
                            data['pagoexamen'] = evidpagoexamen = \
                                inscripcioncohorte.evidenciapagoexamen_set.filter(status=True)[0]
                            if evidpagoexamen.estadorevision == 2:
                                bloqueasubidapago = 2
                    data['bloqueasubidapago'] = bloqueasubidapago
                    data['tienepagoexamen'] = tienepagoexamen
                    permisorequisito = False
                    if inscripcioncohorte.cohortes.fechafinrequisito >= hoy:
                        permisorequisito = True
                    data['permisorequisito'] = permisorequisito
                    if 'comercia' in request.GET:
                        requisitosexcluir = ClaseRequisito.objects.values_list('requisito__id').filter(clasificacion=1,
                                                                                                       status=True)
                        data['requisitos'] = RequisitosMaestria.objects.filter(cohorte=inscripcioncohorte.cohortes,
                                                                               status=True).exclude(
                            requisito__in=requisitosexcluir).order_by('id')
                        data['comercia'] = True
                    else:
                        if InscripcionCohorte.objects.filter(pk=int(encrypt(request.GET['idinscripcioncohorte'])),
                                                             grupo__isnull=True, status=True):
                            requisitosexcluir = ClaseRequisito.objects.values_list('requisito__id').filter(
                                clasificacion=3, status=True)
                            data['requisitos'] = RequisitosMaestria.objects.filter(cohorte=inscripcioncohorte.cohortes,
                                                                                   status=True).exclude(
                                requisito__in=requisitosexcluir).order_by("id")
                        else:
                            gruporequisitos = RequisitosGrupoCohorte.objects.values_list('requisito_id',
                                                                                         flat=True).filter(
                                grupo=inscripcioncohorte.grupo, status=True)
                            data['requisitos'] = RequisitosMaestria.objects.filter(cohorte=inscripcioncohorte.cohortes,
                                                                                   requisito_id__in=gruporequisitos,
                                                                                   status=True).order_by("id")
                    if inscripcioncohorte.integrantegrupoexamenmsc_set.filter(grupoexamen__estado_emailentrevista=2,
                                                                              inscripcion__status=True,
                                                                              status=True).exists():
                        data['integranteexamen'] = inscripcioncohorte.integrantegrupoexamenmsc_set.get(
                            grupoexamen__estado_emailentrevista=2, status=True)
                        ventanaactiva = 2
                    if inscripcioncohorte.integrantegrupoentrevitamsc_set.filter(
                            grupoentrevista__estado_emailentrevista=2, status=True, inscripcion__status=True).exists():
                        data['integranteentrevista'] = inscripcioncohorte.integrantegrupoentrevitamsc_set.get(
                            grupoentrevista__estado_emailentrevista=2, status=True)
                        ventanaactiva = 3
                    if inscripcioncohorte.integrantegrupoentrevitamsc_set.filter(estado_emailadmitido=2,
                                                                                 cohorteadmitidasinproceso__isnull=True,
                                                                                 status=True,
                                                                                 inscripcion__status=True).exists():
                        data['aspiranteadmitido'] = inscripcioncohorte.integrantegrupoentrevitamsc_set.get(
                            estado_emailadmitido=2, cohorteadmitidasinproceso__isnull=True, status=True)
                        ventanaactiva = 4
                        data['otracohorte'] = 0
                    if inscripcioncohorte.integrantegrupoentrevitamsc_set.filter(
                            cohorteadmitidasinproceso__isnull=False, status=True, inscripcion__status=True).exists():
                        data['aspiranteadmitido'] = inscripcioncohorte.integrantegrupoentrevitamsc_set.get(
                            cohorteadmitidasinproceso__isnull=False, status=True)
                        ventanaactiva = 4
                        data['otracohorte'] = 1
                    data['ventanaactiva'] = ventanaactiva
                    return render(request, "adm_admision/listadorequisitosinscripcion.html", data)
                except Exception as ex:
                    pass

            if action == 'aprobarrequisitos':
                try:
                    data['title'] = u'Aprobar Requisitos'
                    data['tipoestado'] = request.GET['tipoestado']
                    data['cant_requisitos'] = request.GET['cant_requisitos']
                    data['inscripcioncohorte'] = InscripcionCohorte.objects.get(
                        pk=int(encrypt(request.GET['aspirante'])), status=True)
                    data['idcohorte'] = int(encrypt(request.GET['idcohorte']))
                    return render(request, "adm_admision/aprobarrequisito.html", data)
                except Exception as ex:
                    pass

            if action == 'delprogramamaestria':
                try:
                    data['title'] = u'Eliminar Cohorte'
                    data['maestria'] = CohorteMaestria.objects.get(pk=int(encrypt(request.GET['idmaestria'])))
                    return render(request, "adm_admision/delprogramamaestria.html", data)
                except Exception as ex:
                    pass

            if action == 'delinscritoentrevista':
                try:
                    data['title'] = u'Eliminar Inscrito'
                    data['iniscritoentrevista'] = IntegranteGrupoEntrevitaMsc.objects.get(
                        pk=int(encrypt(request.GET['idinscrito'])))
                    return render(request, "adm_admision/delinscritoentrevista.html", data)
                except Exception as ex:
                    pass

            if action == 'delrequisitomaestria':
                try:
                    data['title'] = u'Eliminar requisito de maestría'
                    data['requisito'] = RequisitosMaestria.objects.get(pk=int(encrypt(request.GET['id'])))
                    return render(request, "adm_admision/delrequisitomaestria.html", data)
                except Exception as ex:
                    pass

            if action == 'delpreguntamaestria':
                try:
                    data['title'] = u'Eliminar pregunta de maestría'
                    data['pregunta'] = PreguntaMaestria.objects.get(pk=int(encrypt(request.GET['idpregunta'])))
                    return render(request, "adm_admision/delpreguntamaestria.html", data)
                except Exception as ex:
                    pass

            if action == 'deltipopregunta':
                try:
                    data['title'] = u'Eliminar tipo pregunta'
                    data['tipopregunta'] = TipoPreguntasPrograma.objects.get(pk=int(encrypt(request.GET['id'])))
                    return render(request, "adm_admision/deltipopregunta.html", data)
                except Exception as ex:
                    pass

            if action == 'editarentrevistador':
                try:
                    data['title'] = u'Planificar entrevistador'
                    data['form'] = EntrevistadorCohorteForm
                    data['grupoentrevista'] = GrupoEntrevistaMsc.objects.get(pk=int(encrypt(request.GET['idgrupo'])))
                    data['id'] = int(encrypt(request.GET['idcohorte']))
                    return render(request, "adm_admision/addentrevistador.html", data)
                except Exception as ex:
                    pass

            if action == 'delpreguntacohorte':
                try:
                    data['title'] = u'Eliminar pregunta'
                    data['pregunta'] = PreguntaMaestria.objects.get(pk=int(encrypt(request.GET['id'])))
                    return render(request, "adm_admision/delpreguntacohorte.html", data)
                except Exception as ex:
                    pass

            if action == 'inscribircarrera':
                try:
                    data['title'] = u'Inscripción de postulante admitido en carrera'
                    data['integrante'] = integrante = IntegranteGrupoEntrevitaMsc.objects.get(
                        pk=int(encrypt(request.GET['id'])))
                    form = InscripcionCarreraForm(initial={'sede': 1,
                                                           'carrera': integrante.inscripcion.cohortes.maestriaadmision.carrera,
                                                           'modalidad': 1,
                                                           'sesion': 8})
                    miscarreras = persona.mis_carreras()
                    data['form'] = form
                    return render(request, "adm_admision/inscribircarrera.html", data)
                except Exception as ex:
                    pass

            if action == 'listarequisito':
                try:
                    data['title'] = u'Requisitos'
                    requisitos = Requisito.objects.filter(status=True).order_by('nombre')
                    search = None
                    ids = None
                    if 's' in request.GET:
                        search = request.GET['s']
                        ss = search.split(' ')
                        if len(ss) == 1:
                            requisitos = requisitos.filter(
                                Q(nombre__icontains=search) | Q(observacion__icontains=search))
                        else:
                            requisitos = requisitos.filter((Q(nombre__icontains=ss[0]) & Q(nombre__icontains=ss[1])) | (
                                    Q(observacion__icontains=ss[0]) & Q(observacion__icontains=ss[1])))
                    if 'id' in request.GET:
                        ids = int(encrypt(request.GET['id']))
                        requisitos = requisitos.filter(pk=int(encrypt(request.GET['id'])))
                    paging = MiPaginador(requisitos, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['page'] = page
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['requisitos'] = page.object_list
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    return render(request, "adm_admision/view_requisito.html", data)
                except Exception as ex:
                    pass

            if action == 'clasificarrequisito':
                try:
                    data['title'] = u'Clasificación de Requisitos'
                    data['clasificacion'] = clasificacion = TipoClasificacionRequisito.objects.filter(
                        status=True).order_by('id')
                    data['claserequisito'] = claserequisito = ClaseRequisito.objects.filter(status=True)
                    return render(request, "adm_admision/clasificacionrequisito.html", data)
                except Exception as ex:
                    pass

            if action == 'tipoclasificacion':
                try:
                    data['title'] = u'Tipo de Clasificación'
                    tipo = TipoClasificacionRequisito.objects.filter(status=True).order_by('id')
                    search = None
                    ids = None
                    if 's' in request.GET:
                        search = request.GET['s']
                        ss = search.split(' ')
                        if len(ss) == 1:
                            tipo = tipo.filter(
                                Q(descripcion__icontains=search))
                        else:
                            tipo = tipo.filter((Q(descripcion__icontains=ss[0]) & Q(descripcion__icontains=ss[1])))
                    if 'id' in request.GET:
                        ids = int(encrypt(request.GET['id']))
                        tipo = tipo.filter(pk=int(encrypt(request.GET['id'])))
                    paging = MiPaginador(tipo, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['page'] = page
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['tipo'] = page.object_list
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    return render(request, "adm_admision/tipoclasificacionreq.html", data)
                except Exception as ex:
                    pass

            if action == 'addtipoclasificacionreq':
                try:
                    data['form2'] = TipoClasificacionReqForm()
                    template = get_template("adm_admision/modal/modalclasificacionrequisito.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'edittipoclasificacionreq':
                try:
                    data['id'] = request.GET['id']
                    data['filtro'] = filtro = TipoClasificacionRequisito.objects.get(pk=int(encrypt(request.GET['id'])))
                    data['form2'] = TipoClasificacionReqForm(initial=model_to_dict(filtro))
                    template = get_template("adm_admision/modal/modalclasificacionrequisito.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'addclasificacionrequisito':
                try:
                    data['form2'] = ClasificacionRequisitoForm()
                    template = get_template("adm_admision/modal/modalclasificacionrequisito.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'editclasificacionrequisito':
                try:
                    data['id'] = request.GET['id']
                    data['filtro'] = filtro = ClaseRequisito.objects.get(pk=request.GET['id'])
                    form2 = ClasificacionRequisitoForm(initial=model_to_dict(filtro))
                    form2.editar()
                    data['form2'] = form2
                    template = get_template("adm_admision/modal/modalclasificacionrequisito.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'edittipopersonarequisito':
                try:
                    data['id'] = request.GET['id']
                    data['filtro'] = filtro = ClaseRequisito.objects.get(pk=request.GET['id'])
                    form2 = TipoPersonaForm(initial=model_to_dict(filtro))
                    form2.editar()
                    data['form2'] = form2
                    template = get_template("adm_admision/modal/tipopersonarequisito.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'delclasificacionrequisito':
                try:
                    data['title'] = u'ELIMINAR CLASIFICACIÓN'
                    data['clasereq'] = ClaseRequisito.objects.get(pk=request.GET['id'])
                    return render(request, 'adm_admision/deleteclasificacionrequisito.html', data)
                except Exception as ex:
                    pass

            if action == 'listarpreinscritos':
                try:
                    data['title'] = u'Preinscritos'
                    persona = request.session['persona']
                    idc = MaestriasAdmision.objects.values_list('carrera_id', flat=True).filter(
                        cohortemaestria__coordinador=persona.id, status=True).distinct()
                    if idc:
                        preinscripcion = PreInscripcion.objects.filter(status=True, persona__status=True,
                                                                       carrera_id__in=idc).order_by(
                            'persona__apellido1', 'persona__apellido2', 'persona__nombres')
                        if len(preinscripcion) > 0:
                            data['carreras'] = Carrera.objects.filter(
                                id__in=preinscripcion.values_list('carrera__id', flat=True))
                            data['permiso'] = 0
                        else:
                            data['carreras'] = 0
                            data['permiso'] = 0
                    else:

                        preinscripcion = PreInscripcion.objects.filter(status=True, persona__status=True).order_by(
                            'persona__apellido1',
                            'persona__apellido2',
                            'persona__nombres')
                    data['carreras'] = Carrera.objects.filter(
                        id__in=PreInscripcion.objects.values_list('carrera_id', flat=True).filter(
                            status=True).distinct())
                    evidencias = EvidenciasMaestrias.objects.filter(status=True,
                                                                    preinscripcion__persona__status=True).order_by(
                        'persona__apellido1', 'persona__apellido2', 'persona__nombres')
                    search = None
                    search2 = None
                    ids = None
                    if 's' in request.GET:
                        search = request.GET['s']
                        ss = search.split(' ')
                        if len(ss) == 1:
                            preinscripcion = preinscripcion.filter(Q(persona__nombres__icontains=search) |
                                                                   Q(persona__apellido1__icontains=search) |
                                                                   Q(persona__apellido2__icontains=search) |
                                                                   Q(persona__cedula__icontains=search) |
                                                                   Q(persona__pasaporte__icontains=search) |
                                                                   Q(carrera_id__nombre__icontains=search)
                                                                   )
                            evidencias = evidencias.filter(Q(preinscripcion__persona__nombres__icontains=search) |
                                                           Q(preinscripcion__persona__apellido1__icontains=search) |
                                                           Q(preinscripcion__persona__apellido2__icontains=search) |
                                                           Q(preinscripcion__persona__cedula__icontains=search) |
                                                           Q(preinscripcion__persona__pasaporte__icontains=search))
                        else:
                            preinscripcion = preinscripcion.filter(Q(persona__apellido1__icontains=ss[0]) &
                                                                   Q(persona__apellido2__icontains=ss[1]))
                            evidencias = evidencias.filter(Q(preinscripcion__persona__apellido1__icontains=ss[0]) &
                                                           Q(preinscripcion__persona__apellido2__icontains=ss[1]))

                    elif 'g' in request.GET:
                        search = request.GET['g']
                        preinscripcion = preinscripcion.filter(Q(carrera_id=search))

                    if 'id' in request.GET:
                        ids = int(encrypt(request.GET['id']))
                        preinscripcion = preinscripcion.filter(pk=int(encrypt(request.GET['id'])))
                        evidencias = evidencias.filter(pk=int(encrypt(request.GET['id'])))
                    paging = MiPaginador(preinscripcion, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['page'] = page
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['preinscripcion'] = page.object_list
                    data['evidencias'] = page.object_list
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""

                    return render(request, "adm_admision/listarpreinscritos.html", data)
                except Exception as ex:
                    pass

            if action == 'listformatocarrerapreins':
                try:
                    data['title'] = u' Configuración Formato Programa'
                    formato = FormatoCarreraIpec.objects.filter(status=True).order_by('carrera__nombre')
                    search = None
                    ids = None
                    if 's' in request.GET:
                        search = request.GET['s']
                        ss = search.split(' ')
                        if len(ss) == 1:
                            formato = formato.filter(carrera__nombre__icontains=search)
                        else:
                            formato = formato.filter(Q(carrera__nombre__icontains=ss[0]))
                    if 'id' in request.GET:
                        ids = int(encrypt(request.GET['id']))
                        formato = formato.filter(pk=int(encrypt(request.GET['id'])))
                    paging = MiPaginador(formato, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['page'] = page
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['formato'] = page.object_list
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    return render(request, "adm_admision/listarformatocarrerapreins.html", data)
                except Exception as ex:
                    pass

            if action == 'editpreinscrito':
                try:
                    data['title'] = u'Editar Preinscrito'
                    data['preinscrito'] = preinscrito = PreInscripcion.objects.get(status=True,
                                                                                   pk=int(encrypt(request.GET['id'])))
                    data['form'] = PreInscripcionForm(initial={'carrera': preinscrito.carrera})
                    return render(request, 'adm_admision/editpreinscrito.html', data)
                except Exception as ex:
                    pass

            if action == 'editformatocarrera':
                try:
                    data['formato'] = formato = FormatoCarreraIpec.objects.get(status=True,
                                                                               pk=int(encrypt(request.GET['id'])))
                    data['id'] = int(encrypt(request.GET['id']))
                    data['form'] = FormatoCarreraForm(
                        initial={'carrera': formato.carrera, 'correomaestria': formato.correomaestria,
                                 'archivo': formato.archivo})
                    template = get_template("adm_admision/editformatocarrera.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, "mensaje": ex})

            if action == 'evidencias':
                try:
                    data['title'] = u'Listado de Evidencias'
                    data['preinscripcion'] = preinscripcion = PreInscripcion.objects.get(
                        pk=int(encrypt(request.GET['id'])))
                    data['evidenciasmaestrias'] = EvidenciasMaestrias.objects.filter(preinscripcion=preinscripcion,
                                                                                     status=True)
                    return render(request, "adm_admision/evidenciaspreinscrito.html", data)
                except Exception as ex:
                    pass

            if action == 'delformatocarrera':
                try:
                    data['title'] = u'Eliminar Formato de Carrera'
                    data['formato'] = FormatoCarreraIpec.objects.get(pk=int(encrypt(request.GET['id'])))
                    return render(request, "adm_admision/delformatocarrera.html", data)
                except Exception as ex:
                    pass

            if action == 'addformatopreinscrito':
                try:
                    form = FormatoCarreraForm()
                    formato = FormatoCarreraIpec.objects.values_list('carrera_id', flat=True).order_by(
                        'carrera__nombre')
                    form.fields['carrera'].queryset = Carrera.objects.filter(status=True, coordinacion=7).exclude(
                        id__in=formato)
                    data['form'] = form
                    template = get_template("adm_admision/addformatoinscrito.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'addobservacionpreinscrito':
                try:
                    data['title'] = u'Agregar Observación'
                    data['observacion'] = observacion = EvidenciasMaestrias.objects.get(
                        preinscripcion_id=int(encrypt(request.GET['id'])))
                    data['form'] = EvidenciasMaestriasForm(initial={'observaciones': observacion.observaciones})
                    return render(request, 'adm_admision/addobservacionpreinscrito.html', data)
                except Exception as ex:
                    pass

            if action == 'listapreguntas':
                try:
                    data['title'] = u'Lista de Preguntas'
                    preguntas = PreguntasPrograma.objects.filter(status=True).order_by('descripcion').order_by(
                        'tipopregunta', 'descripcion')
                    search = None
                    ids = None
                    if 's' in request.GET:
                        search = request.GET['s']
                        ss = search.split(' ')
                        if len(ss) == 1:
                            preguntas = preguntas.filter(descripcion__icontains=search)
                        else:
                            preguntas = preguntas.filter(
                                (Q(descripcion__icontains=ss[0]) & Q(descripcion__icontains=ss[1])))
                    if 'id' in request.GET:
                        ids = int(encrypt(request.GET['id']))
                        preguntas = preguntas.filter(pk=int(encrypt(request.GET['id'])))
                    paging = MiPaginador(preguntas, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['page'] = page
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['preguntas'] = page.object_list
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    return render(request, "adm_admision/view_preguntas.html", data)
                except Exception as ex:
                    pass

            if action == 'listatipopreguntas':
                try:
                    data['title'] = u'Lista de Tipo Preguntas'
                    tipopreguntas = TipoPreguntasPrograma.objects.filter(status=True).order_by('descripcion')
                    search = None
                    ids = None
                    if 's' in request.GET:
                        search = request.GET['s']
                        ss = search.split(' ')
                        if len(ss) == 1:
                            tipopreguntas = tipopreguntas.filter(descripcion__icontains=search)
                        else:
                            tipopreguntas = tipopreguntas.filter(
                                (Q(descripcion__icontains=ss[0]) & Q(descripcion__icontains=ss[1])))
                    if 'id' in request.GET:
                        ids = int(encrypt(request.GET['id']))
                        tipopreguntas = tipopreguntas.filter(pk=int(encrypt(request.GET['id'])))
                    paging = MiPaginador(tipopreguntas, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['page'] = page
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['tipopreguntas'] = page.object_list
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    return render(request, "adm_admision/view_tipopreguntas.html", data)
                except Exception as ex:
                    pass

            if action == 'addrequisito':
                try:
                    form = RequisitoForm()
                    data['form'] = form
                    template = get_template("adm_admision/addrequisito.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'addpregunta':
                try:
                    data['title'] = u'Adicionar Pregunta'
                    form = PreguntaForm()
                    data['form'] = form
                    return render(request, 'adm_admision/addpregunta.html', data)
                except Exception as ex:
                    pass

            if action == 'editpregunta':
                try:
                    data['title'] = u'Editar Pregunta'
                    data['pregunta'] = pregunta = PreguntasPrograma.objects.get(status=True,
                                                                                pk=int(encrypt(request.GET['id'])))
                    data['form'] = PreguntaForm(
                        initial={'descripcion': pregunta.descripcion, 'tipopregunta': pregunta.tipopregunta})
                    return render(request, 'adm_admision/editpregunta.html', data)
                except Exception as ex:
                    pass

            if action == 'delpregunta':
                try:
                    data['title'] = u'Eliminar pregunta'
                    data['pregunta'] = PreguntasPrograma.objects.get(pk=int(encrypt(request.GET['id'])))
                    return render(request, "adm_admision/delpregunta.html", data)
                except Exception as ex:
                    pass

            if action == 'addtipopregunta':
                try:
                    data['title'] = u'Adicionar Tipo Pregunta'
                    form = TipoPreguntaForm()
                    data['form'] = form
                    return render(request, 'adm_admision/addtipopregunta.html', data)
                except Exception as ex:
                    pass

            if action == 'edittipopregunta':
                try:
                    data['title'] = u'Editar Tipo Pregunta'
                    data['tipopregunta'] = tipopregunta = TipoPreguntasPrograma.objects.get(status=True, pk=int(
                        encrypt(request.GET['id'])))
                    data['form'] = TipoPreguntaForm(initial={'descripcion': tipopregunta.descripcion})
                    return render(request, 'adm_admision/edittipopregunta.html', data)
                except Exception as ex:
                    pass

            if action == 'listarpreinscritos':
                try:
                    data['title'] = u'Lista de Pre inscritos'
                    data['preinscritos'] = PreInscripcion.objects.filter(status=True)

                    return render(request, 'adm_admision/listarpreinscritos.html', data)
                except Exception as ex:
                    pass

            if action == 'editrequisito':
                try:
                    data['requisito'] = requisito = Requisito.objects.get(status=True,
                                                                          pk=int(encrypt(request.GET['id'])))
                    data['id'] = int(encrypt(request.GET['id']))
                    data['form'] = RequisitoForm(initial={'nombre': requisito.nombre,
                                                          'observacion': requisito.observacion,
                                                          'tipoarchivo': requisito.tipoarchivo,
                                                          'activo': requisito.activo})
                    # datos para seleccionar en combo de clasificacion
                    listaclase = []
                    clases = requisito.claserequisito_set.filter(status=True)
                    for c in clases:
                        listaclase.append(c.clasificacion.id)
                    data['clasificacion'] = listaclase
                    template = get_template("adm_admision/editrequisito.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'delrequisito':
                try:
                    data['title'] = u'Eliminar requisito'
                    data['requisito'] = Requisito.objects.get(pk=int(encrypt(request.GET['id'])))
                    return render(request, "adm_admision/delrequisito.html", data)
                except Exception as ex:
                    pass

            if action == 'delpreinscrito':
                try:
                    data['title'] = u'Eliminar preinscrito'
                    data['preinscrito'] = PreInscripcion.objects.get(pk=int(encrypt(request.GET['id'])))
                    return render(request, "adm_admision/deletepreinscripcion.html", data)
                except Exception as ex:
                    pass

            if action == 'addgrupoexamen':
                try:
                    data['cohorte'] = CohorteMaestria.objects.get(status=True, pk=int(encrypt(request.GET['id'])))
                    form = GrupoExamenForm()
                    form.fields['profesor'].queryset = Profesor.objects.none()
                    data['form'] = form
                    template = get_template("adm_admision/addgrupoexamen.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'editgrupoexamen':
                try:
                    data['cohorte'] = cohorte = CohorteMaestria.objects.get(status=True,
                                                                            pk=int(encrypt(request.GET['idcohorte'])))
                    data['grupo'] = grupo = GrupoExamenMsc.objects.get(status=True, pk=int(encrypt(request.GET['id'])),
                                                                       cohorte=cohorte)
                    if grupo.profesor:
                        data['tiene_profesor'] = True
                        form = GrupoExamenForm(initial={'profesor': grupo.profesor if grupo.profesor else 0,
                                                        'lugar': grupo.lugar,
                                                        'fecha': grupo.fecha,
                                                        'hora': grupo.hora,
                                                        'visible': grupo.visible,
                                                        'urlzoom': grupo.urlzoom,
                                                        'observacion': grupo.observacion})
                        if grupo.profesor:
                            form.fields['profesor'].queryset = Profesor.objects.filter(pk=grupo.profesor.pk)
                        else:
                            form.fields['profesor'].queryset = Profesor.objects.none()
                    else:
                        data['tiene_profesor'] = False
                        form = GrupoExamenForm()
                    data['form'] = form
                    template = get_template("adm_admision/editgrupoexamen.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'delintegrante':
                try:
                    data['title'] = u'Eliminar Integrante'
                    data['t'] = int(request.GET['t'])
                    data['integrante'] = IntegranteGrupoExamenMsc.objects.get(pk=int(encrypt(request.GET['idi'])),
                                                                              grupoexamen_id=int(
                                                                                  encrypt(request.GET['idg'])))
                    return render(request, "adm_admision/delintegrante.html", data)
                except Exception as ex:
                    pass

            if action == 'delgrupoexamen':
                try:
                    data['title'] = u'Eliminar grupo de Examen'
                    data['grupo'] = GrupoExamenMsc.objects.get(pk=int(encrypt(request.GET['id'])),
                                                               cohorte_id=int(encrypt(request.GET['idc'])))
                    return render(request, "adm_admision/delgrupoexamen.html", data)
                except Exception as ex:
                    pass

            if action == 'creargrupomoodle':
                try:
                    data['title'] = u'Crear grupo de Examen moodle'
                    data['grupo'] = GrupoExamenMsc.objects.get(pk=int(encrypt(request.GET['id'])),
                                                               cohorte_id=int(encrypt(request.GET['idc'])))
                    return render(request, "adm_admision/creargrupomoodle.html", data)
                except Exception as ex:
                    pass

            if action == 'grupoexamen':
                try:
                    data['title'] = u'Grupos de examen'
                    data['cohorte'] = cohorte = CohorteMaestria.objects.get(status=True,
                                                                            pk=int(encrypt(request.GET['id'])))
                    data['inscritosaprobados'] = cohorte.inscripcioncohorte_set.values('id').filter(status=True,
                                                                                                    activo=True,
                                                                                                    estado_emailevidencia=2,
                                                                                                    aproboproceso=False,
                                                                                                    integrantegrupoexamenmsc__isnull=True,
                                                                                                    inscripcion__status=True).count()
                    data['inscritosnonotificados'] = cohorte.inscripcioncohorte_set.values('id').filter(
                        estado_emailevidencia=1, estado_aprobador=2, status=True, inscripcion__status=True).count()
                    data['inscritosenproceso'] = cohorte.inscripcioncohorte_set.values('id').filter(
                        estado_emailevidencia=1, estado_aprobador=1, status=True, inscripcion__status=True).count()
                    data['inscritossingrupos'] = cohorte.inscripcioncohorte_set.filter(estado_aprobador=2, status=True,
                                                                                       integrantegrupoexamenmsc__isnull=True).count()
                    data['grupos'] = cohorte.grupoexamenmsc_set.filter(status=True).order_by('id')

                    return render(request, "adm_admision/grupoexamen.html", data)
                except Exception as ex:
                    pass

            if action == 'notasmoodle':
                try:
                    data['title'] = u'Notas de moodle'
                    search = None
                    data['grupo'] = grupo = GrupoExamenMsc.objects.get(pk=int(encrypt(request.GET['id'])),
                                                                       cohorte_id=int(encrypt(request.GET['idc'])))
                    data['utiliza_validacion_calificaciones'] = variable_valor('UTILIZA_VALIDACION_CALIFICACIONES')
                    data['habilitado_ingreso_calificaciones'] = True
                    # if persona.id == grupo.profesor.persona.id:
                    #     data['habilitado_ingreso_calificaciones'] = True
                    if 's' in request.GET:
                        search = request.GET['s']
                        ss = search.split(' ')
                        inscritos = grupo.integrantegrupoexamenmsc_set.filter(
                            Q(inscripcion__inscripcionaspirante__persona__nombres__icontains=search) |
                            Q(inscripcion__inscripcionaspirante__persona__apellido1__icontains=search) |
                            Q(inscripcion__inscripcionaspirante__persona__apellido2__icontains=search) |
                            Q(inscripcion__inscripcionaspirante__persona__cedula__icontains=search) |
                            Q(inscripcion__inscripcionaspirante__persona__pasaporte__icontains=search) |
                            Q(inscripcion__inscripcionaspirante__persona__usuario__username__icontains=search),
                            status=True, inscripcion__status=True).distinct()
                    else:
                        inscritos = grupo.integrantegrupoexamenmsc_set.filter(status=True,
                                                                              inscripcion__status=True).order_by(
                            'inscripcion__inscripcionaspirante__persona__apellido1',
                            'inscripcion__inscripcionaspirante__persona__apellido2')
                    numerofilas = 30
                    paging = MiPaginador(inscritos, numerofilas)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                            if p == 1:
                                numerofilasguiente = numerofilas
                            else:
                                numerofilasguiente = numerofilas * (p - 1)
                        else:
                            p = paginasesion
                            if p == 1:
                                numerofilasguiente = numerofilas
                            else:
                                numerofilasguiente = numerofilas * (p - 1)
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['page'] = page
                    data['numerofilasguiente'] = numerofilasguiente
                    data['numeropagina'] = p
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['inscritos'] = page.object_list
                    data['search'] = search if search else ""
                    return render(request, "adm_admision/notasmoodle.html", data)
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            elif action == 'extraernotasmoodle':
                try:
                    data['title'] = u'Importar calificaciones de moodle'
                    data['grupo'] = grupo = GrupoExamenMsc.objects.get(pk=int(encrypt(request.GET['id'])))
                    return render(request, "adm_admision/extraernotasmoodle.html", data)
                except Exception as ex:
                    pass

            elif action == 'copiaraentrevista':
                try:
                    data['title'] = u'Copiar calificaciones a entrevista'
                    data['grupoexamen'] = grupo = GrupoExamenMsc.objects.get(pk=int(encrypt(request.GET['id'])))
                    if GrupoEntrevistaMsc.objects.filter(cohortes=grupo.cohorte, status=True).exists():
                        data['grupoentrevista'] = GrupoEntrevistaMsc.objects.filter(cohortes=grupo.cohorte, status=True)
                        template = get_template("adm_admision/copiaraentrevista.html")
                        return JsonResponse({"result": "ok", "data": template.render(data)})
                    else:
                        return JsonResponse({"result": "bad", "mensaje": u"No posee grupo de entrevista activo."})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            if action == 'listaintegrante':
                try:
                    data['title'] = u'Lista de Integrantes'
                    data['grupo'] = grupo = GrupoExamenMsc.objects.get(pk=int(encrypt(request.GET['id'])),
                                                                       cohorte_id=int(encrypt(request.GET['idc'])),
                                                                       status=True)
                    data['integrantes'] = grupo.integrantegrupoexamenmsc_set.filter(status=True,
                                                                                    inscripcion__status=True)
                    return render(request, "adm_admision/view_integrantes.html", data)
                except Exception as ex:
                    pass

            if action == 'addtablapronderacion':
                try:
                    data['title'] = u'Adicionar Tabla ponderación'
                    if 'id' in request.GET:
                        data['cohorte'] = CohorteMaestria.objects.get(status=True, pk=int(encrypt(request.GET['id'])))
                    data['p'] = int(request.GET['p'])
                    data['form'] = TablaEntrevistaMaestriaForm()
                    return render(request, "adm_admision/addtablaponderacion.html", data)
                except Exception as ex:
                    pass

            if action == 'edittablapronderacion':
                try:
                    data['title'] = u'Editar Tabla ponderación'
                    if 'idc' in request.GET:
                        data['cohorte'] = CohorteMaestria.objects.get(status=True, pk=int(encrypt(request.GET['idc'])))
                    data['p'] = int(request.GET['p'])
                    data['tabla'] = tabla = TablaEntrevistaMaestria.objects.get(status=True,
                                                                                pk=int(encrypt(request.GET['id'])))
                    data['form'] = TablaEntrevistaMaestriaForm(initial={'nombre': tabla.nombre})
                    return render(request, "adm_admision/edittablaponderacion.html", data)
                except Exception as ex:
                    pass

            if action == 'deltabla':
                try:
                    data['title'] = u'Eliminar tabla de ponderación'
                    data['tabla'] = TablaEntrevistaMaestria.objects.get(status=True, pk=int(encrypt(request.GET['id'])))
                    return render(request, "adm_admision/deltabla.html", data)
                except Exception as ex:
                    pass

            if action == 'deldetalletabla':
                try:
                    data['title'] = u'Eliminar es estado de la tabla de ponderación'
                    data['detalle'] = EstadoEntrevista.objects.get(status=True, pk=int(encrypt(request.GET['id'])))
                    return render(request, "adm_admision/deldetalletabla.html", data)
                except Exception as ex:
                    pass

            if action == 'tablaponderacion':
                try:
                    data['title'] = u'Ponderación'
                    tablas = TablaEntrevistaMaestria.objects.filter(status=True).order_by('nombre')
                    search = None
                    ids = None
                    if 's' in request.GET:
                        search = request.GET['s']
                        ss = search.split(' ')
                        if len(ss) == 1:
                            tablas = tablas.filter(Q(nombre__icontains=search))
                        else:
                            tablas = tablas.filter((Q(nombre__icontains=ss[0]) & Q(nombre__icontains=ss[1])))
                    if 'id' in request.GET:
                        ids = int(encrypt(request.GET['id']))
                        tablas = tablas.filter(pk=int(encrypt(request.GET['id'])))
                    paging = MiPaginador(tablas, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['page'] = page
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['tablas'] = page.object_list
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    return render(request, "adm_admision/tablaponderacion.html", data)
                except Exception as ex:
                    pass

            if action == 'formatoexamen':
                try:
                    __author__ = 'Unemi'
                    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on',
                                      num_format_str='#,##0.00')
                    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                    title = easyxf(
                        'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    style1 = easyxf(num_format_str='D-MMM-YY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('exp_xls_post_part')
                    # ws.write_merge(0, 0, 0, 7, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=listado_' + random.randint(1,
                                                                                                       10000).__str__() + '.xls'
                    row_num = 0

                    columns = [
                        (u"CEDULA", 3000),
                        (u"APELLIDOS", 6000),
                        (u"NOMBRES", 6000),
                        (u"EMAIL", 6000),
                        (u"NOTA1", 2500),
                        (u"NOTA2", 2500),
                        (u"TOTAL", 2500),
                    ]
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    date_format = xlwt.XFStyle()
                    date_format.num_format_str = 'yyyy/mm/dd'

                    row_num = 0
                    listado = IntegranteGrupoExamenMsc.objects.filter(
                        grupoexamen__cohorte_id=int(encrypt(request.GET['idcohorte'])), status=True,
                        inscripcion__status=True).order_by('inscripcion__inscripcionaspirante__persona__apellido1',
                                                           'inscripcion__inscripcionaspirante__persona__apellido2',
                                                           'inscripcion__inscripcionaspirante__persona__nombres')
                    for lista in listado:
                        row_num += 1
                        campo1 = lista.inscripcion.inscripcionaspirante.persona.cedula
                        campo2 = lista.inscripcion.inscripcionaspirante.persona.apellido1 + ' ' + lista.inscripcion.inscripcionaspirante.persona.apellido2
                        campo3 = lista.inscripcion.inscripcionaspirante.persona.nombres
                        campo4 = lista.inscripcion.inscripcionaspirante.persona.email
                        if not lista.notaexa:
                            campo5 = 0
                        else:
                            campo5 = lista.notaexa
                        if not lista.notatest:
                            campo6 = 0
                        else:
                            campo6 = lista.notatest
                        if not lista.notafinal:
                            campo7 = 0
                        else:
                            campo7 = lista.notafinal
                        ws.write(row_num, 0, campo1, font_style2)
                        ws.write(row_num, 1, campo2, font_style2)
                        ws.write(row_num, 2, campo3, font_style2)
                        ws.write(row_num, 3, campo4, font_style2)
                        ws.write(row_num, 4, campo5, font_style2)
                        ws.write(row_num, 5, campo6, font_style2)
                        ws.write(row_num, 6, campo7, font_style2)
                    wb.save(response)
                    return response
                except Exception as ex:
                    pass

            if action == 'descargarlistadoinscritos':
                try:
                    __author__ = 'Unemi'
                    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on',
                                      num_format_str='#,##0.00')
                    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                    title = easyxf(
                        'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    style1 = easyxf(num_format_str='D-MMM-YY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('exp_xls_post_part')
                    # ws.write_merge(0, 0, 0, 7, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=listado_inscritos' + random.randint(1,
                                                                                                                10000).__str__() + '.xls'
                    row_num = 0
                    columns = [
                        (u"N.", 2000),
                        (u"CEDULA", 4000),
                        (u"APELLIDO 1", 6000),
                        (u"APELLIDO 2", 6000),
                        (u"NOMBRE", 6000),
                        (u"USUARIO", 4000),
                        (u"EMAIL", 6000),
                        (u"TELEFONO", 4000),
                        (u"TOTAL DOCUMENTOS SUBIDOS", 2000),
                        (u"TOTAL DOCUMENTOS", 2000),
                        (u"TOTAL DOCUMENTOS RECHAZADOS", 2000),
                        (u"FECHA INSCRIPCIÓN", 4000),
                        (u"SOLICITUD DE BECA", 15000),
                        (u"FECHA APROBACION", 3000),
                        (u"FECHA NOTIFICACIÓN", 3000),
                        (u"USUARIOS SUBIERON EVIDENCIA", 8000),
                        (u"USUARIOS MODIFICARON EVIDENCIA", 8000),
                    ]
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    date_format = xlwt.XFStyle()
                    date_format.num_format_str = 'yyyy/mm/dd'
                    listado = InscripcionCohorte.objects.filter(cohortes_id=int(encrypt(request.GET['idcohorte'])),
                                                                status=True,
                                                                estado_aprobador=request.GET['idestado']).order_by(
                        'inscripcionaspirante__persona__apellido1', 'inscripcionaspirante__persona__apellido2')
                    row_num = 0
                    for lista in listado:
                        row_num += 1
                        nombreusuario = '-'
                        campo2 = '-'
                        if lista.inscripcionaspirante.persona.usuario:
                            if not lista.inscripcionaspirante.persona.es_personalactivo():
                                nombreusuario = lista.inscripcionaspirante.persona.usuario.username
                        if lista.inscripcionaspirante.persona.cedula:
                            campo2 = lista.inscripcionaspirante.persona.cedula
                        else:
                            if lista.inscripcionaspirante.persona.pasaporte:
                                campo2 = lista.inscripcionaspirante.persona.pasaporte
                        campo3 = lista.inscripcionaspirante.persona.apellido1
                        campo4 = lista.inscripcionaspirante.persona.apellido2
                        campo5 = lista.inscripcionaspirante.persona.nombres
                        campo9 = lista.inscripcionaspirante.persona.email
                        campo10 = lista.inscripcionaspirante.persona.telefono
                        campo11 = lista.fecha_aprobador
                        campo12 = lista.fecha_emailevidencia
                        campo13 = ''
                        listausuariocreacion = ""
                        listausuariomodificacion = ""
                        listadoevidenciasubidas = lista.evidenciarequisitosaspirante_set.filter(status=True)
                        for levid in listadoevidenciasubidas:
                            listausuariocreacion = listausuariocreacion + levid.usuario_creacion.username + ','
                            if levid.usuario_modificacion:
                                listausuariomodificacion = listausuariomodificacion + levid.usuario_modificacion.username + ','
                        campo14 = str(lista.total_evidencias())
                        campo15 = str(lista.total_evidenciasgrupocohorte())
                        campo16 = str(lista.total_evidenciasrechazadas())
                        if lista.tipobeca:
                            campo13 = lista.tipobeca.descuentoposgrado.nombre
                        campo17 = lista.fecha_creacion
                        ws.write(row_num, 0, row_num, font_style2)
                        ws.write(row_num, 1, campo2, font_style2)
                        ws.write(row_num, 2, campo3, font_style2)
                        ws.write(row_num, 3, campo4, font_style2)
                        ws.write(row_num, 4, campo5, font_style2)
                        ws.write(row_num, 5, nombreusuario, font_style2)
                        ws.write(row_num, 6, campo9, font_style2)
                        ws.write(row_num, 7, campo10, font_style2)
                        ws.write(row_num, 8, campo14, font_style2)
                        ws.write(row_num, 9, campo15, font_style2)
                        ws.write(row_num, 10, campo16, font_style2)
                        ws.write(row_num, 11, campo17, date_format)
                        ws.write(row_num, 12, campo13, font_style2)
                        ws.write(row_num, 13, campo11, date_format)
                        ws.write(row_num, 14, campo12, date_format)
                        ws.write(row_num, 15, listausuariocreacion, font_style2)
                        ws.write(row_num, 16, listausuariomodificacion, font_style2)
                    wb.save(response)
                    return response
                except Exception as ex:
                    pass

            if action == 'descargarlistadoadmitidos':
                try:
                    __author__ = 'Unemi'
                    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on',
                                      num_format_str='#,##0.00')
                    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                    title = easyxf(
                        'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    style1 = easyxf(num_format_str='D-MMM-YY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('exp_xls_post_part')
                    # ws.write_merge(0, 0, 0, 7, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    response = HttpResponse(content_type="application/ms-excel")
                    cohorte = CohorteMaestria.objects.get(pk=int(encrypt(request.GET['idcohorte'])))
                    response[
                        'Content-Disposition'] = 'attachment; filename=' + cohorte.descripcion + '_' + cohorte.maestriaadmision.carrera.alias + random.randint(
                        1, 10000).__str__() + '.xls'

                    row_num = 0
                    columns = [
                        (u"N.", 2000),
                        (u"CEDULA", 4000),
                        (u"NOMBRES", 10000),
                        (u"EMAIL", 6000),
                        (u"TELEFONO", 4000),
                        (u"RUBRO GENERADO", 4000),
                        (u"RUBRO CANCELADO", 4000),
                        (u"VALOR", 4000),
                        (u"SALDO", 4000),
                        (u"RUBRO", 4000),
                        (u"FECHA ASIGNACION RUBRO", 4000),
                        (u"FECHA ABONO RUBRO", 4000),
                        (u"DIAS", 4000),
                    ]
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    date_format = xlwt.XFStyle()
                    date_format.num_format_str = 'yyyy/mm/dd'

                    listadoadmitidossinproceso = IntegranteGrupoEntrevitaMsc.objects.filter(estado_emailadmitido=2,
                                                                                            cohorteadmitidasinproceso__isnull=True,
                                                                                            grupoentrevista__cohortes=cohorte,
                                                                                            status=True,
                                                                                            inscripcion__status=True).order_by(
                        'inscripcion__inscripcionaspirante__persona__apellido1',
                        'inscripcion__inscripcionaspirante__persona__apellido2')
                    row_num = 0
                    for lista in listadoadmitidossinproceso:
                        row_num += 1
                        if lista.inscripcion.inscripcionaspirante.persona.cedula:
                            campo2 = lista.inscripcion.inscripcionaspirante.persona.cedula
                        if lista.inscripcion.inscripcionaspirante.persona.pasaporte:
                            campo2 = lista.inscripcion.inscripcionaspirante.persona.pasaporte
                        campo5 = lista.inscripcion.inscripcionaspirante.persona.apellido1 + ' ' + lista.inscripcion.inscripcionaspirante.persona.apellido2 + ' ' + lista.inscripcion.inscripcionaspirante.persona.nombres
                        campo9 = lista.inscripcion.inscripcionaspirante.persona.email
                        campo10 = lista.inscripcion.inscripcionaspirante.persona.telefono
                        campo11 = 'NO'
                        campo12 = 'NO'
                        valorrubro = 0
                        valorsaldo = 0
                        idrubro = 0
                        fechacreacionrubro = None
                        fechaabonorubro = None
                        if lista.inscripcion.tipocobro == 2:
                            if lista.inscripcion.genero_rubro_matricula():
                                campo11 = 'SI'
                            if lista.inscripcion.cancelo_rubro_matricula():
                                campo12 = 'SI'

                            if Rubro.objects.filter(status=True, admisionposgradotipo=2, inscripcion=lista.inscripcion):
                                fechacreacionrubro = Rubro.objects.filter(status=True, admisionposgradotipo=2,
                                                                          inscripcion=lista.inscripcion)[0]
                                idrubro = fechacreacionrubro.id
                                if fechacreacionrubro.pago_set.filter(status=True):
                                    rpago = fechacreacionrubro.pago_set.filter(status=True)[0]
                                    fechaabonorubro = rpago.fecha
                                valorrubro = Rubro.objects.filter(status=True, admisionposgradotipo=2,
                                                                  inscripcion=lista.inscripcion).aggregate(
                                    valor=Sum('valor'))['valor']
                                valorsaldo = Rubro.objects.filter(status=True, admisionposgradotipo=2,
                                                                  inscripcion=lista.inscripcion).aggregate(
                                    valor=Sum('saldo'))['valor']

                        if lista.inscripcion.tipocobro == 3:
                            if lista.inscripcion.genero_rubro_programa():
                                campo11 = 'SI'
                            if lista.inscripcion.cancelo_rubro_programa():
                                campo12 = 'SI'

                            if Rubro.objects.filter(status=True, admisionposgradotipo=3, inscripcion=lista.inscripcion):
                                fechacreacionrubro = Rubro.objects.filter(status=True, admisionposgradotipo=3,
                                                                          inscripcion=lista.inscripcion)[0]
                                idrubro = fechacreacionrubro.id
                                if fechacreacionrubro.pago_set.filter(status=True):
                                    rpago = fechacreacionrubro.pago_set.filter(status=True)[0]
                                    fechaabonorubro = rpago.fecha
                                valorrubro = Rubro.objects.filter(status=True, admisionposgradotipo=3,
                                                                  inscripcion=lista.inscripcion).aggregate(
                                    valor=Sum('valor'))['valor']
                                valorsaldo = Rubro.objects.filter(status=True, admisionposgradotipo=3,
                                                                  inscripcion=lista.inscripcion).aggregate(
                                    valor=Sum('saldo'))['valor']
                        diasresta = 0
                        ws.write(row_num, 0, row_num, font_style2)
                        ws.write(row_num, 1, campo2, font_style2)
                        ws.write(row_num, 2, campo5, font_style2)
                        ws.write(row_num, 3, campo9, font_style2)
                        ws.write(row_num, 4, campo10, font_style2)
                        ws.write(row_num, 5, campo11, font_style2)
                        ws.write(row_num, 6, campo12, font_style2)
                        ws.write(row_num, 7, valorrubro, font_style2)
                        ws.write(row_num, 8, valorsaldo, font_style2)
                        ws.write(row_num, 9, idrubro, font_style2)
                        if fechacreacionrubro:
                            ws.write(row_num, 10, fechacreacionrubro.fecha_creacion, date_format)
                        else:
                            ws.write(row_num, 10, '', font_style2)
                        if fechaabonorubro:
                            ws.write(row_num, 11, fechaabonorubro, date_format)
                        else:
                            ws.write(row_num, 11, '', font_style2)
                        if fechacreacionrubro and fechaabonorubro:
                            fechacreacionrubro = datetime.strptime(str(fechacreacionrubro.fecha_creacion.date()),
                                                                   "%Y-%m-%d")
                            fechaabonorubro = datetime.strptime(str(fechaabonorubro), "%Y-%m-%d")
                            diasresta = fechaabonorubro - fechacreacionrubro
                            ws.write(row_num, 12, diasresta.days, font_style2)
                    wb.save(response)
                    return response
                except Exception as ex:
                    pass

            if action == 'descargarlistadoadmitidoscohorte':
                try:
                    __author__ = 'Unemi'
                    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on',
                                      num_format_str='#,##0.00')
                    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                    title = easyxf(
                        'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    style1 = easyxf(num_format_str='D-MMM-YY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('exp_xls_post_part')
                    # ws.write_merge(0, 0, 0, 7, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    response = HttpResponse(content_type="application/ms-excel")
                    cohorte = CohorteMaestria.objects.get(pk=int(encrypt(request.GET['idcohorte'])))
                    response[
                        'Content-Disposition'] = 'attachment; filename=' + cohorte.descripcion + '_' + cohorte.maestriaadmision.carrera.alias + random.randint(
                        1, 10000).__str__() + '.xls'

                    row_num = 0
                    columns = [
                        (u"N.", 2000),
                        (u"CEDULA", 4000),
                        (u"NOMBRES", 10000),
                        (u"EMAIL", 6000),
                        (u"TELEFONO", 4000),
                        (u"RUBRO GENERADO", 4000),
                        (u"RUBRO CANCELADO", 4000),
                        (u"VALOR", 4000),
                        (u"SALDO", 4000),
                        (u"RUBRO", 4000),
                        (u"FECHA ASIGNACION RUBRO", 4000),
                        (u"FECHA ABONO RUBRO", 4000),
                        (u"DIAS", 4000),
                    ]
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    date_format = xlwt.XFStyle()
                    date_format.num_format_str = 'yyyy/mm/dd'
                    if cohorte.tipo == 2:
                        listados = InscripcionCohorte.objects.filter(cohortes=cohorte,cohortes__tipo=2,estado_aprobador=2, integrantegrupoexamenmsc__estado=2, status=True)
                    if cohorte.tipo == 3:
                        listados = InscripcionCohorte.objects.filter(cohortes=cohorte,estado_aprobador=2,status=True)
                        #listadoadmitidossinproceso = InscripcionCohorte.objects.filter(cohortes=cohorte, cohortes__aplica=False, estado_aprobador=2, status=True).order_by('inscripcionaspirante__persona__apellido1', 'inscripcionaspirante__persona__apellido2')
                    linstadoinscripcion = InscripcionCohorte.objects.filter(cohortes=cohorte, tipocobro__in=[2,3], status=True)
                    listadoadmitidossinproceso = listados | linstadoinscripcion
                    row_num = 0
                    idrubro = 0
                    for lista in listadoadmitidossinproceso:
                        fechaabonorubro = None
                        fechacreacionrubro = None
                        row_num += 1
                        if lista.inscripcionaspirante.persona.cedula:
                            campo2 = lista.inscripcionaspirante.persona.cedula
                        if lista.inscripcionaspirante.persona.pasaporte:
                            campo2 = lista.inscripcionaspirante.persona.pasaporte
                        campo5 = lista.inscripcionaspirante.persona.apellido1 + ' ' + lista.inscripcionaspirante.persona.apellido2 + ' ' + lista.inscripcionaspirante.persona.nombres
                        campo9 = lista.inscripcionaspirante.persona.email
                        campo10 = lista.inscripcionaspirante.persona.telefono
                        campo11 = 'NO'
                        campo12 = 'NO'
                        valorrubro = 0
                        valorsaldo = 0
                        if lista.tipocobro == 2:
                            if lista.genero_rubro_matricula():
                                campo11 = 'SI'
                            if lista.cancelo_rubro_matricula():
                                campo12 = 'SI'

                            if Rubro.objects.filter(status=True, admisionposgradotipo=2, inscripcion=lista):
                                fechacreacionrubro = Rubro.objects.filter(status=True, admisionposgradotipo=2, inscripcion=lista)[0]
                                idrubro = fechacreacionrubro.id
                                if fechacreacionrubro.pago_set.filter(status=True):
                                    rpago = fechacreacionrubro.pago_set.filter(status=True)[0]
                                    fechaabonorubro = rpago.fecha
                                valorrubro = Rubro.objects.filter(status=True, admisionposgradotipo=2,inscripcion=lista).aggregate(valor=Sum('valor'))['valor']
                                valorsaldo = Rubro.objects.filter(status=True, admisionposgradotipo=2,inscripcion=lista).aggregate(valor=Sum('saldo'))['valor']

                        if lista.tipocobro == 3:
                            if lista.genero_rubro_programa():
                                campo11 = 'SI'
                            if lista.cancelo_rubro_programa():
                                campo12 = 'SI'

                            if Rubro.objects.filter(status=True, admisionposgradotipo=3, inscripcion=lista):
                                fechacreacionrubro = Rubro.objects.filter(status=True, admisionposgradotipo=3, inscripcion=lista)[0]
                                idrubro = fechacreacionrubro.id
                                if fechacreacionrubro.pago_set.filter(status=True):
                                    rpago = fechacreacionrubro.pago_set.filter(status=True)[0]
                                    fechaabonorubro = rpago.fecha
                                valorrubro = Rubro.objects.filter(status=True, admisionposgradotipo=3,inscripcion=lista).aggregate(valor=Sum('valor'))['valor']
                                valorsaldo = Rubro.objects.filter(status=True, admisionposgradotipo=3,inscripcion=lista).aggregate(valor=Sum('saldo'))['valor']

                        ws.write(row_num, 0, row_num, font_style2)
                        ws.write(row_num, 1, campo2, font_style2)
                        ws.write(row_num, 2, campo5, font_style2)
                        ws.write(row_num, 3, campo9, font_style2)
                        ws.write(row_num, 4, campo10, font_style2)
                        ws.write(row_num, 5, campo11, font_style2)
                        ws.write(row_num, 6, campo12, font_style2)
                        ws.write(row_num, 7, valorrubro, font_style2)
                        ws.write(row_num, 8, valorsaldo, font_style2)
                        ws.write(row_num, 9, idrubro, font_style2)
                        if fechacreacionrubro:
                            ws.write(row_num, 10, fechacreacionrubro.fecha_creacion, date_format)
                        else:
                            ws.write(row_num, 10, '', font_style2)
                        if fechaabonorubro:
                            ws.write(row_num, 11, fechaabonorubro, date_format)
                        else:
                            ws.write(row_num, 11, '', font_style2)
                        if fechacreacionrubro and fechaabonorubro:
                            fechacreacionrubro = datetime.strptime(str(fechacreacionrubro.fecha_creacion.date()),
                                                                   "%Y-%m-%d")
                            fechaabonorubro = datetime.strptime(str(fechaabonorubro), "%Y-%m-%d")
                            diasresta = fechaabonorubro - fechacreacionrubro
                            ws.write(row_num, 12, diasresta.days, font_style2)
                    wb.save(response)
                    return response
                except Exception as ex:
                    pass

            if action == 'descargarlistadocomprobantes':
                try:
                    __author__ = 'Unemi'
                    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on',num_format_str='#,##0.00')
                    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                    title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    style1 = easyxf(num_format_str='D-MMM-YY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('exp_xls_post_part')
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=comprobantes' + random.randint(1, 10000).__str__() + '.xls'
                    row_num = 0
                    columns = [
                        (u"N.", 2000),
                        (u"FECHA REGISTRO", 4000),
                        (u"DIAS VALIDAR REGISTRO", 4000),
                        (u"CEDULA", 4000),
                        (u"NOMBRES", 10000),
                        (u"EMAIL", 6000),
                        (u"TELEFONO", 4000),
                        (u"CARRERA", 4000),
                        (u"CURSO", 4000),
                        (u"OBSERVACION", 4000),
                        (u"FECHA PAGO", 4000),
                        (u"BANCO DESTINO", 4000),
                        (u"TIPO TRANSACCION", 4000),
                        (u"VALOR", 4000),
                        (u"ESTADO", 4000),
                    ]
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    date_format = xlwt.XFStyle()
                    date_format.num_format_str = 'yyyy/mm/dd'
                    listadocomprobantes = ComprobanteAlumno.objects.filter(status=True)
                    for lista in listadocomprobantes:
                        row_num += 1
                        campo1 = lista.fecha_creacion
                        campo2 = ''

                        if lista.persona.cedula:
                            campo3 = lista.persona.cedula
                        if lista.persona.pasaporte:
                            campo3 = lista.persona.pasaporte
                        campo4 = lista.persona.apellido1 + ' ' + lista.persona.apellido2 + ' ' + lista.persona.nombres
                        campo5 = lista.persona.email
                        campo6 = lista.persona.telefono
                        campo7 = lista.carrera
                        campo8 = lista.curso
                        campo9 = lista.observacion
                        campo10 = lista.fechapago
                        campo11 = lista.cuentadeposito
                        campo12 = lista.get_tipocomprobante_display()
                        campo13 = lista.valor
                        campo14 = lista.get_estados_display()

                        ws.write(row_num, 0, row_num, font_style2)
                        ws.write(row_num, 1, campo1, date_format)
                        if lista.fecha_creacion and lista.fechavalidacion:
                            fechacreacioncomprobante = datetime.strptime(str(lista.fecha_creacion.date()), "%Y-%m-%d")
                            fechavalida = datetime.strptime(str(lista.fechavalidacion), "%Y-%m-%d")
                            diasresta = fechavalida - fechacreacioncomprobante
                            ws.write(row_num, 2, diasresta.days, font_style2)
                        else:
                            ws.write(row_num, 2, 0, font_style2)
                        ws.write(row_num, 3, campo3, font_style2)
                        ws.write(row_num, 4, campo4, font_style2)
                        ws.write(row_num, 5, campo5, font_style2)
                        ws.write(row_num, 6, campo6, font_style2)
                        ws.write(row_num, 7, campo7, font_style2)
                        ws.write(row_num, 8, campo8, font_style2)
                        ws.write(row_num, 9, campo9, font_style2)
                        ws.write(row_num, 10, campo10, date_format)
                        ws.write(row_num, 11, campo11, font_style2)
                        ws.write(row_num, 12, campo12, font_style2)
                        ws.write(row_num, 13, campo13, font_style2)
                        ws.write(row_num, 14, campo14, font_style2)
                    wb.save(response)
                    return response
                except Exception as ex:
                    pass

            if action == 'descargarlistadoaprobados':
                try:
                    __author__ = 'Unemi'
                    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on',
                                      num_format_str='#,##0.00')
                    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                    title = easyxf(
                        'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    style1 = easyxf(num_format_str='D-MMM-YY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('exp_xls_post_part')
                    # ws.write_merge(0, 0, 0, 7, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=listado_' + random.randint(1,
                                                                                                       10000).__str__() + '.xls'
                    row_num = 0
                    listado = CuotaPago.objects.filter(
                        pago__inscripcioncohorte__cohortes_id=int(encrypt(request.GET['idcohorte'])), status=True)

                    columns = [
                        (u"N.", 2000),
                        (u"CEDULA", 6000),
                        (u"APELLIDO 1", 6000),
                        (u"APELLIDO 2", 6000),
                        (u"NOMBRE", 6000),

                        (u"VALOR $", 2500),
                        (u"FECHA PAGO", 3000),
                        (u"PROGRAMA", 15000),
                        (u"COHORTE", 15000),
                    ]
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    date_format = xlwt.XFStyle()
                    date_format.num_format_str = 'yyyy/mm/dd'

                    row_num = 0
                    for lista in listado:
                        row_num += 1
                        campo2 = lista.pago.inscripcioncohorte.inscripcionaspirante.persona.cedula
                        campo3 = lista.pago.inscripcioncohorte.inscripcionaspirante.persona.apellido1
                        campo4 = lista.pago.inscripcioncohorte.inscripcionaspirante.persona.apellido2
                        campo5 = lista.pago.inscripcioncohorte.inscripcionaspirante.persona.nombres
                        campo6 = lista.valor
                        campo7 = lista.fechapago
                        campo8 = lista.pago.inscripcioncohorte.cohortes.maestriaadmision.carrera.nombre
                        campo9 = lista.pago.inscripcioncohorte.cohortes.descripcion
                        ws.write(row_num, 0, row_num, font_style2)
                        ws.write(row_num, 1, campo2, font_style2)
                        ws.write(row_num, 2, campo3, font_style2)
                        ws.write(row_num, 3, campo4, font_style2)
                        ws.write(row_num, 4, campo5, font_style2)
                        ws.write(row_num, 5, campo6, font_style2)
                        ws.write(row_num, 6, campo7, date_format)
                        ws.write(row_num, 7, campo8, font_style2)
                        ws.write(row_num, 8, campo9, font_style2)
                    wb.save(response)
                    return response
                except Exception as ex:
                    pass

            if action == 'descargarpostulates':
                try:
                    __author__ = 'Unemi'
                    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on',
                                      num_format_str='#,##0.00')
                    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                    title = easyxf(
                        'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    style1 = easyxf(num_format_str='D-MMM-YY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('exp_xls_post_part')
                    # ws.write_merge(0, 0, 0, 7, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=listado_postulantes' + random.randint(1,
                                                                                                                  10000).__str__() + '.xls'
                    row_num = 0
                    columns = [
                        (u"N.", 2000),
                        (u"CEDULA", 4000),
                        (u"APELLIDO 1", 6000),
                        (u"APELLIDO 2", 6000),
                        (u"NOMBRE", 6000),
                        (u"CANTON", 10000),
                        (u"EMAIL", 6000),
                        (u"TELEFONO", 4000),
                        (u"SEXO", 4000),
                    ]
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    date_format = xlwt.XFStyle()
                    date_format.num_format_str = 'yyyy/mm/dd'
                    listado = InscripcionAspirante.objects.filter(status=True).order_by('persona__apellido1',
                                                                                        'persona__apellido2')
                    row_num = 0
                    for lista in listado:
                        row_num += 1
                        if lista.persona.cedula:
                            campo2 = lista.persona.cedula
                        else:
                            campo2 = lista.persona.pasaporte
                        campo3 = lista.persona.apellido1
                        campo4 = lista.persona.apellido2
                        campo5 = lista.persona.nombres
                        campo8 = ''
                        if lista.persona.canton:
                            campo8 = lista.persona.canton.nombre
                        campo9 = lista.persona.email
                        campo10 = lista.persona.telefono
                        campo11 = lista.persona.sexo.nombre

                        ws.write(row_num, 0, row_num, font_style2)
                        ws.write(row_num, 1, campo2, font_style2)
                        ws.write(row_num, 2, campo3, font_style2)
                        ws.write(row_num, 3, campo4, font_style2)
                        ws.write(row_num, 4, campo5, font_style2)
                        ws.write(row_num, 5, campo8, font_style2)
                        ws.write(row_num, 6, campo9, font_style2)
                        ws.write(row_num, 7, campo10, font_style2)
                        ws.write(row_num, 8, campo11, font_style2)
                    wb.save(response)
                    return response
                except Exception as ex:
                    pass

            if action == 'reportexcelpreinscritos':
                try:
                    __author__ = 'Unemi'
                    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on',
                                      num_format_str='#,##0.00')
                    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                    title = easyxf(
                        'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    style1 = easyxf(num_format_str='D-MMM-YY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False

                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('exp_xls_post_part')
                    # ws.write_merge(0, 0, 0, 7, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=listado_preinscrito' + random.randint(1,
                                                                                                                  10000).__str__() + '.xls'
                    row_num = 0
                    columns = [
                        (u"N.", 2000),
                        (u"NOMBRES", 5000),
                        (u"APELLIDOS", 6000),
                        (u"CÉDULA", 4000),
                        (u"PROGRAMA", 16000),
                        (u"EMAIL", 9000),
                        (u"ESTADO", 4000),
                        (u"OBSERVACIÓN", 4000),
                    ]
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    date_format = xlwt.XFStyle()
                    date_format.num_format_str = 'yyyy/mm/dd'
                    listado = EvidenciasMaestrias.objects.filter(preinscripcion__status=True,
                                                                 preinscripcion_id__status=True).order_by(
                        'preinscripcion__persona__apellido1', 'preinscripcion__persona__apellido2')
                    row_num = 0
                    for lista in listado:
                        row_num += 1
                        campo2 = lista.preinscripcion.persona.nombres
                        campo3 = lista.preinscripcion.persona.apellido1 + ' ' + lista.preinscripcion.persona.apellido2
                        if lista.preinscripcion.persona.cedula:
                            campo4 = lista.preinscripcion.persona.cedula
                        else:
                            campo4 = lista.preinscripcion.persona.pasaporte
                        campo5 = lista.preinscripcion.carrera.nombre
                        campo6 = lista.preinscripcion.persona.email
                        if lista.preinscripcion.aceptarpreinscripcion == True:
                            campo7 = 'Aceptado'
                        else:
                            campo7 = 'No aceptado'

                        if lista.observaciones:
                            campo8 = lista.observaciones
                        else:
                            campo8 = ''

                        ws.write(row_num, 0, row_num, font_style2)
                        ws.write(row_num, 1, campo2, font_style2)
                        ws.write(row_num, 2, campo3, font_style2)
                        ws.write(row_num, 3, campo4, font_style2)
                        ws.write(row_num, 4, campo5, font_style2)
                        ws.write(row_num, 5, campo6, font_style2)
                        ws.write(row_num, 6, campo7, font_style2)
                        ws.write(row_num, 7, campo8, font_style2)
                    wb.save(response)
                    return response
                except Exception as ex:
                    pass

            if action == 'reportexcelpreinscritoscarrera':
                try:
                    __author__ = 'Unemi'
                    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on',
                                      num_format_str='#,##0.00')
                    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                    title = easyxf(
                        'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    style1 = easyxf(num_format_str='D-MMM-YY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False

                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('exp_xls_post_part')
                    # ws.write_merge(0, 0, 0, 7, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=listado_preinscrito' + random.randint(1,
                                                                                                                  10000).__str__() + '.xls'
                    row_num = 0
                    columns = [
                        (u"N.", 2000),
                        (u"NOMBRES", 5000),
                        (u"APELLIDOS", 6000),
                        (u"CÉDULA", 4000),
                        (u"PROGRAMA", 16000),
                        (u"EMAIL", 9000),
                        (u"ESTADO", 4000),
                        (u"DOCUMENTOS", 4000),
                    ]
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    date_format = xlwt.XFStyle()
                    date_format.num_format_str = 'yyyy/mm/dd'
                    persona = request.session['persona']
                    idc = MaestriasAdmision.objects.values_list('carrera_id', flat=True).filter(
                        cohortemaestria__coordinador=persona.id, status=True).distinct()
                    preinscritos = PreInscripcion.objects.filter(status=True).order_by('persona__apellido1',
                                                                                       'persona__apellido2')
                    if idc:
                        preinscritos = preinscritos.filter(carrera=idc)
                        if int(request.GET['id']) > 0:
                            listado = preinscritos.filter(carrera_id=int(request.GET['id']))
                    else:
                        listado = preinscritos.filter(carrera_id=int(request.GET['id']))
                        if len(listado) == 0:
                            listado = preinscritos.filter(status=True)
                    row_num = 0
                    for lista in listado:
                        row_num += 1
                        campo2 = lista.persona.nombres
                        campo3 = lista.persona.apellido1 + ' ' + lista.persona.apellido2
                        if lista.persona.cedula:
                            campo4 = lista.persona.cedula
                        else:
                            campo4 = lista.persona.pasaporte
                        campo5 = lista.carrera.nombre
                        campo6 = lista.persona.email
                        if lista.aceptarpreinscripcion == True:
                            campo7 = 'Aceptado'
                        else:
                            campo7 = 'No aceptado'

                        if lista.evidencias == True:
                            campo8 = 'Con documentos'
                        else:
                            campo8 = 'Sin documentos'

                        ws.write(row_num, 0, row_num, font_style2)
                        ws.write(row_num, 1, campo2, font_style2)
                        ws.write(row_num, 2, campo3, font_style2)
                        ws.write(row_num, 3, campo4, font_style2)
                        ws.write(row_num, 4, campo5, font_style2)
                        ws.write(row_num, 5, campo6, font_style2)
                        ws.write(row_num, 6, campo7, font_style2)
                        ws.write(row_num, 7, campo8, font_style2)

                    wb.save(response)
                    return response
                except Exception as ex:
                    pass

            if action == 'descargarlistadogrupos':
                try:
                    __author__ = 'Unemi'
                    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on',
                                      num_format_str='#,##0.00')
                    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                    title = easyxf(
                        'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    style1 = easyxf(num_format_str='D-MMM-YY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('exp_xls_post_part')
                    # ws.write_merge(0, 0, 0, 7, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=listado_postulantes' + random.randint(1,
                                                                                                                  10000).__str__() + '.xls'
                    row_num = 0
                    columns = [
                        (u"N.", 2000),
                        (u"LUGAR", 7000),
                        (u"FECHA", 4000),
                        (u"HORA", 4000),
                        (u"OBSERVACION", 7000),
                        (u"CEDULA", 4000),
                        (u"APELLIDO 1", 6000),
                        (u"APELLIDO 2", 6000),
                        (u"NOMBRE", 6000),
                        (u"EMAIL", 6000),
                        (u"TELEFONO", 4000),
                        (u"NOTAFINAL", 4000),
                        (u"ESTADO", 4000),
                    ]
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    date_format = xlwt.XFStyle()
                    hora_format = xlwt.XFStyle()
                    date_format.num_format_str = 'yyyy/mm/dd'
                    hora_format = xlwt.easyxf(num_format_str='h:mm')
                    data['cohorte'] = cohorte = CohorteMaestria.objects.get(status=True,
                                                                            pk=int(encrypt(request.GET['idcohorte'])))
                    listado = IntegranteGrupoExamenMsc.objects.filter(grupoexamen__cohorte=cohorte, status=True,
                                                                      grupoexamen__cohorte__status=True,
                                                                      inscripcion__status=True).order_by(
                        'grupoexamen__id', 'inscripcion__inscripcionaspirante__persona__apellido1')
                    row_num = 0
                    for lista in listado:
                        row_num += 1
                        if lista.inscripcion.inscripcionaspirante.persona.cedula:
                            campo2 = lista.inscripcion.inscripcionaspirante.persona.cedula
                        else:
                            campo2 = lista.inscripcion.inscripcionaspirante.persona.pasaporte
                        campo3 = lista.inscripcion.inscripcionaspirante.persona.apellido1
                        campo4 = lista.inscripcion.inscripcionaspirante.persona.apellido2
                        campo5 = lista.inscripcion.inscripcionaspirante.persona.nombres
                        # campo8 = ''
                        # if lista.inscripcion.inscripcionaspirante.persona.canton:
                        #     campo8 = lista.inscripcion.inscripcionaspirante.persona.canton.nombre
                        campo9 = lista.inscripcion.inscripcionaspirante.persona.email
                        campo10 = lista.inscripcion.inscripcionaspirante.persona.telefono
                        # campo11 = lista.inscripcion.inscripcionaspirante.persona.sexo.nombre
                        campo12 = lista.grupoexamen.lugar
                        campo13 = lista.grupoexamen.fecha
                        campo14 = lista.grupoexamen.hora
                        campo15 = lista.grupoexamen.observacion
                        campo16 = lista.notafinal
                        campo17 = lista.get_estado_display()
                        ws.write(row_num, 0, row_num, font_style2)
                        ws.write(row_num, 1, campo12, font_style2)
                        ws.write(row_num, 2, campo13, date_format)
                        ws.write(row_num, 3, campo14.__str__(), font_style2)
                        ws.write(row_num, 4, campo15, font_style2)
                        ws.write(row_num, 5, campo2, font_style2)
                        ws.write(row_num, 6, campo3, font_style2)
                        ws.write(row_num, 7, campo4, font_style2)
                        ws.write(row_num, 8, campo5, font_style2)
                        # ws.write(row_num, 9, campo8, font_style2)
                        ws.write(row_num, 9, campo9, font_style2)
                        ws.write(row_num, 10, campo10, font_style2)
                        # ws.write(row_num, 12, campo11, font_style2)
                        ws.write(row_num, 11, campo16, font_style2)
                        ws.write(row_num, 12, campo17, font_style2)
                    wb.save(response)
                    return response
                except Exception as ex:
                    pass

            if action == 'descargarlistadoporgrupos':
                try:
                    __author__ = 'Unemi'
                    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on',
                                      num_format_str='#,##0.00')
                    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                    title = easyxf(
                        'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    style1 = easyxf(num_format_str='D-MMM-YY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('exp_xls_post_part')
                    # ws.write_merge(0, 0, 0, 7, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=listado_postulantes' + random.randint(1,
                                                                                                                  10000).__str__() + '.xls'
                    row_num = 0
                    columns = [
                        (u"N.", 2000),
                        (u"LUGAR", 7000),
                        (u"FECHA", 4000),
                        (u"HORA", 4000),
                        (u"OBSERVACION", 7000),
                        (u"CEDULA", 4000),
                        (u"APELLIDO 1", 6000),
                        (u"APELLIDO 2", 6000),
                        (u"NOMBRE", 6000),
                        (u"EMAIL", 6000),
                        (u"TELEFONO", 4000),
                        (u"NOTAFINAL", 4000),
                        (u"ESTADO", 4000),
                    ]
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    date_format = xlwt.XFStyle()
                    hora_format = xlwt.XFStyle()
                    date_format.num_format_str = 'yyyy/mm/dd'
                    hora_format = xlwt.easyxf(num_format_str='h:mm')
                    listado = IntegranteGrupoExamenMsc.objects.filter(
                        grupoexamen_id=int(encrypt(request.GET['idgrupo'])), status=True,
                        inscripcion__status=True).order_by('inscripcion__inscripcionaspirante__persona__apellido1')
                    row_num = 0
                    for lista in listado:
                        row_num += 1
                        if lista.inscripcion.inscripcionaspirante.persona.cedula:
                            campo2 = lista.inscripcion.inscripcionaspirante.persona.cedula
                        else:
                            campo2 = lista.inscripcion.inscripcionaspirante.persona.pasaporte
                        campo3 = lista.inscripcion.inscripcionaspirante.persona.apellido1
                        campo4 = lista.inscripcion.inscripcionaspirante.persona.apellido2
                        campo5 = lista.inscripcion.inscripcionaspirante.persona.nombres
                        # campo8 = ''
                        # if lista.inscripcion.inscripcionaspirante.persona.canton:
                        #     campo8 = lista.inscripcion.inscripcionaspirante.persona.canton.nombre
                        campo9 = lista.inscripcion.inscripcionaspirante.persona.email
                        campo10 = lista.inscripcion.inscripcionaspirante.persona.telefono
                        # campo11 = lista.inscripcion.inscripcionaspirante.persona.sexo.nombre
                        campo12 = lista.grupoexamen.lugar
                        campo13 = lista.grupoexamen.fecha
                        campo14 = lista.grupoexamen.hora
                        campo15 = lista.grupoexamen.observacion
                        campo16 = lista.notafinal
                        campo17 = lista.get_estado_display()
                        ws.write(row_num, 0, row_num, font_style2)
                        ws.write(row_num, 1, campo12, font_style2)
                        ws.write(row_num, 2, campo13, date_format)
                        ws.write(row_num, 3, campo14.__str__(), font_style2)
                        ws.write(row_num, 4, campo15, font_style2)
                        ws.write(row_num, 5, campo2, font_style2)
                        ws.write(row_num, 6, campo3, font_style2)
                        ws.write(row_num, 7, campo4, font_style2)
                        ws.write(row_num, 8, campo5, font_style2)
                        # ws.write(row_num, 9, campo8, font_style2)
                        ws.write(row_num, 9, campo9, font_style2)
                        ws.write(row_num, 10, campo10, font_style2)
                        # ws.write(row_num, 12, campo11, font_style2)
                        ws.write(row_num, 11, campo16, font_style2)
                        ws.write(row_num, 12, campo17, font_style2)
                    wb.save(response)
                    return response
                except Exception as ex:
                    pass

            if action == 'descargarlistadogruposentrevista':
                try:
                    __author__ = 'Unemi'
                    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on',
                                      num_format_str='#,##0.00')
                    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                    title = easyxf(
                        'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    style1 = easyxf(num_format_str='D-MMM-YY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('exp_xls_post_part')
                    # ws.write_merge(0, 0, 0, 7, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=listado_postulantes' + random.randint(1,
                                                                                                                  10000).__str__() + '.xls'
                    row_num = 0
                    columns = [
                        (u"N.", 2000),
                        (u"LUGAR", 7000),
                        (u"FECHA", 4000),
                        (u"HORA", 4000),
                        (u"OBSERVACION", 7000),
                        (u"CEDULA", 4000),
                        (u"APELLIDO 1", 6000),
                        (u"APELLIDO 2", 6000),
                        (u"NOMBRE", 6000),
                        (u"CANTON", 10000),
                        (u"EMAIL", 6000),
                        (u"TELEFONO", 4000),
                        (u"SEXO", 4000),
                    ]
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    date_format = xlwt.XFStyle()
                    hora_format = xlwt.XFStyle()
                    date_format.num_format_str = 'yyyy/mm/dd'
                    hora_format = xlwt.easyxf(num_format_str='h:mm')
                    data['cohorte'] = cohorte = CohorteMaestria.objects.get(status=True,
                                                                            pk=int(encrypt(request.GET['idcohorte'])))
                    listado = IntegranteGrupoEntrevitaMsc.objects.filter(grupoentrevista__cohortes=cohorte, status=True,
                                                                         grupoentrevista__cohortes__status=True,
                                                                         inscripcion__status=True).order_by(
                        'grupoentrevista__id', 'inscripcion__inscripcionaspirante__persona__apellido1')
                    row_num = 0
                    for lista in listado:
                        row_num += 1
                        if lista.inscripcion.inscripcionaspirante.persona.cedula:
                            campo2 = lista.inscripcion.inscripcionaspirante.persona.cedula
                        else:
                            campo2 = lista.inscripcion.inscripcionaspirante.persona.pasaporte
                        campo3 = lista.inscripcion.inscripcionaspirante.persona.apellido1
                        campo4 = lista.inscripcion.inscripcionaspirante.persona.apellido2
                        campo5 = lista.inscripcion.inscripcionaspirante.persona.nombres
                        campo8 = ''
                        if lista.inscripcion.inscripcionaspirante.persona.canton:
                            campo8 = lista.inscripcion.inscripcionaspirante.persona.canton.nombre
                        campo9 = lista.inscripcion.inscripcionaspirante.persona.email
                        campo10 = lista.inscripcion.inscripcionaspirante.persona.telefono
                        campo11 = ''
                        if lista.inscripcion.inscripcionaspirante.persona.sexo:
                            campo11 = lista.inscripcion.inscripcionaspirante.persona.sexo.nombre
                        campo12 = lista.grupoentrevista.lugar
                        campo13 = lista.grupoentrevista.fecha
                        campo14 = lista.grupoentrevista.horainicio
                        campo15 = lista.grupoentrevista.observacion
                        ws.write(row_num, 0, row_num, font_style2)
                        ws.write(row_num, 1, campo12, font_style2)
                        ws.write(row_num, 2, campo13, date_format)
                        ws.write(row_num, 3, campo14.__str__(), font_style2)
                        ws.write(row_num, 4, campo15, font_style2)
                        ws.write(row_num, 5, campo2, font_style2)
                        ws.write(row_num, 6, campo3, font_style2)
                        ws.write(row_num, 7, campo4, font_style2)
                        ws.write(row_num, 8, campo5, font_style2)
                        ws.write(row_num, 9, campo8, font_style2)
                        ws.write(row_num, 10, campo9, font_style2)
                        ws.write(row_num, 11, campo10, font_style2)
                        ws.write(row_num, 12, campo11, font_style2)
                    wb.save(response)
                    return response
                except Exception as ex:
                    pass

            if action == 'descargarlistadoporgruposentrevista':
                try:
                    __author__ = 'Unemi'
                    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on',
                                      num_format_str='#,##0.00')
                    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                    title = easyxf(
                        'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    style1 = easyxf(num_format_str='D-MMM-YY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('exp_xls_post_part')
                    # ws.write_merge(0, 0, 0, 7, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=listado_postulantes' + random.randint(1,
                                                                                                                  10000).__str__() + '.xls'
                    row_num = 0
                    columns = [
                        (u"N.", 2000),
                        (u"LUGAR", 7000),
                        (u"FECHA", 4000),
                        (u"HORA", 4000),
                        (u"OBSERVACION", 7000),
                        (u"CEDULA", 4000),
                        (u"APELLIDO 1", 6000),
                        (u"APELLIDO 2", 6000),
                        (u"NOMBRE", 6000),
                        (u"CANTON", 10000),
                        (u"EMAIL", 6000),
                        (u"TELEFONO", 4000),
                        (u"SEXO", 4000),
                    ]
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    date_format = xlwt.XFStyle()
                    hora_format = xlwt.XFStyle()
                    date_format.num_format_str = 'yyyy/mm/dd'
                    hora_format = xlwt.easyxf(num_format_str='h:mm')
                    listado = IntegranteGrupoEntrevitaMsc.objects.filter(
                        grupoentrevista_id=int(encrypt(request.GET['idgrupo'])), status=True,
                        inscripcion__status=True).order_by('inscripcion__inscripcionaspirante__persona__apellido1')
                    row_num = 0
                    for lista in listado:
                        row_num += 1
                        if lista.inscripcion.inscripcionaspirante.persona.cedula:
                            campo2 = lista.inscripcion.inscripcionaspirante.persona.cedula
                        else:
                            campo2 = lista.inscripcion.inscripcionaspirante.persona.pasaporte
                        campo3 = lista.inscripcion.inscripcionaspirante.persona.apellido1
                        campo4 = lista.inscripcion.inscripcionaspirante.persona.apellido2
                        campo5 = lista.inscripcion.inscripcionaspirante.persona.nombres
                        campo8 = ''
                        if lista.inscripcion.inscripcionaspirante.persona.canton:
                            campo8 = lista.inscripcion.inscripcionaspirante.persona.canton.nombre
                        campo9 = lista.inscripcion.inscripcionaspirante.persona.email
                        campo10 = lista.inscripcion.inscripcionaspirante.persona.telefono
                        campo11 = ''
                        if lista.inscripcion.inscripcionaspirante.persona.sexo:
                            campo11 = lista.inscripcion.inscripcionaspirante.persona.sexo.nombre
                        campo12 = lista.grupoentrevista.lugar
                        campo13 = lista.grupoentrevista.fecha
                        campo14 = lista.grupoentrevista.horainicio
                        campo15 = lista.grupoentrevista.observacion
                        ws.write(row_num, 0, row_num, font_style2)
                        ws.write(row_num, 1, campo12, font_style2)
                        ws.write(row_num, 2, campo13, date_format)
                        ws.write(row_num, 3, campo14.__str__(), font_style2)
                        ws.write(row_num, 4, campo15, font_style2)
                        ws.write(row_num, 5, campo2, font_style2)
                        ws.write(row_num, 6, campo3, font_style2)
                        ws.write(row_num, 7, campo4, font_style2)
                        ws.write(row_num, 8, campo5, font_style2)
                        ws.write(row_num, 9, campo8, font_style2)
                        ws.write(row_num, 10, campo9, font_style2)
                        ws.write(row_num, 11, campo10, font_style2)
                        ws.write(row_num, 12, campo11, font_style2)
                    wb.save(response)
                    return response
                except Exception as ex:
                    pass

            elif action == 'resetear':
                try:
                    data['title'] = u'Resetear clave del usuario'
                    data['inscripcion'] = InscripcionCohorte.objects.get(pk=request.GET['id'])
                    return render(request, "adm_admision/resetear.html", data)
                except Exception as ex:
                    pass

            elif action == 'emailrecordatorio':
                try:
                    data['title'] = u'Email recordatorio'
                    data['inscripcion'] = InscripcionCohorte.objects.get(pk=request.GET['id'])
                    return render(request, "adm_admision/emailrecordatorio.html", data)
                except Exception as ex:
                    pass

            elif action == 'regresaproceso':
                try:
                    data['title'] = u'Regresar a estado de proceso'
                    data['inscripcion'] = InscripcionCohorte.objects.get(pk=request.GET['id'])
                    return render(request, "adm_admision/regresaproceso.html", data)
                except Exception as ex:
                    pass

            if action == 'listatablas':
                try:
                    data['title'] = u'Selección de tabla de ponderación para el cohorte'
                    data['cohorte'] = CohorteMaestria.objects.get(status=True, pk=int(encrypt(request.GET['id'])))
                    # data['tablas'] = TablaEntrevistaMaestria.objects.filter(status=True)
                    tablas = TablaEntrevistaMaestria.objects.filter(status=True).order_by('nombre')
                    search = None
                    ids = None
                    if 's' in request.GET:
                        search = request.GET['s']
                        ss = search.split(' ')
                        if len(ss) == 1:
                            tablas = tablas.filter(Q(nombre__icontains=search))
                        else:
                            tablas = tablas.filter((Q(nombre__icontains=ss[0]) & Q(nombre__icontains=ss[1])))
                    paging = MiPaginador(tablas, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['page'] = page
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['tablas'] = page.object_list
                    data['search'] = search if search else ""
                    data['idc'] = request.GET['id']
                    return render(request, "adm_admision/selecciontabla.html", data)
                except Exception as ex:
                    pass

            elif action == 'generarmatricula':
                try:
                    data['title'] = u'Generar matricula'
                    data['integrante'] = IntegranteGrupoEntrevitaMsc.objects.get(pk=encrypt(request.GET['id']))
                    return render(request, "adm_admision/generarmatricula.html", data)
                except Exception as ex:
                    pass

            elif action == 'configuracionfirma':
                try:
                    q = None
                    ids = None
                    if 's' in request.GET:
                        q = request.GET['s'].upper().strip()
                        s = q.split(" ")
                        if len(s) == 1:
                            per = Administrativo.objects.filter(
                                (Q(persona__nombres__icontains=q) | Q(persona__apellido1__icontains=q) |
                                 Q(persona__cedula__icontains=q) | Q(persona__apellido2__icontains=q) |
                                 Q(persona__cedula__contains=q)), Q(persona__status=True)
                            ).distinct()[:15]
                        elif len(s) == 2:
                            per = Administrativo.objects.filter(
                                (Q(persona__apellido1__contains=s[0]) & Q(persona__apellido2__contains=s[1])) |
                                (Q(persona__nombres__icontains=s[0]) & Q(persona__nombres__icontains=s[1])) |
                                (Q(persona__nombres__icontains=s[0]) & Q(persona__apellido1__contains=s[1]))
                            ).filter(status=True, persona__status=True).distinct()[:15]
                        else:
                            per = Administrativo.objects.filter(
                                (Q(persona__nombres__contains=s[0]) & Q(persona__apellido1__contains=s[1]) &
                                 Q(persona__apellido2__contains=s[2])) | (Q(persona__nombres__contains=s[0]) &
                                                                          Q(persona__nombres__contains=s[1]) & Q(
                                            persona__apellido1__contains=s[2]))
                            ).filter(status=True, persona__status=True).distinct()[:15]

                        conf = ConfigurarFirmaAdmisionPosgrado.objects.filter(status=True,
                                                                              administrativo_id__in=per.values_list(
                                                                                  'id')).order_by('id')
                    else:
                        conf = ConfigurarFirmaAdmisionPosgrado.objects.filter(status=True).order_by('id')
                    paging = MiPaginador(conf, 15)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['title'] = u'Configurar firma'
                    data['paging'] = paging
                    data['page'] = page
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['search'] = q if q else ""
                    data['ids'] = ids if ids else ""
                    data['listadoConfigFirma'] = page.object_list
                    return render(request, "adm_admision/configuracionfirma.html", data)
                except Exception as ex:
                    pass


            elif action == 'addfirma':
                try:
                    data['title'] = u'Agregar firma'
                    f = ConfigurarFirmaAdmisionPosgradoForm()
                    if 'id' in request.GET:
                        lista2, firma = {}, ''
                        administrativo = Administrativo.objects.get(persona_id=int(request.GET['id']))
                        if administrativo.persona.tiene_firma():
                            listafirma = administrativo.persona.tiene_firma()
                            if listafirma[0][1] == 2:
                                firma = str(listafirma[0][2])
                        lista2['nombre'] = administrativo.persona.nombre_titulos3y4()
                        lista2['firma'] = firma
                        return JsonResponse({"result": "ok", "data": lista2})

                    data['form'] = f
                    return render(request, "adm_admision/addfirma.html", data)
                except Exception as ex:
                    pass

            elif action == 'editfirma':
                try:
                    data['title'] = u'Editar firma'
                    f = ConfigurarFirmaAdmisionPosgradoForm()
                    idFirma = int(encrypt(request.GET['id']))
                    configFirma = ConfigurarFirmaAdmisionPosgrado.objects.get(pk=idFirma)
                    data['id'] = configFirma.id
                    data['idf'] = int(encrypt(request.GET['idf']))
                    data['nombre'] = configFirma.administrativo.persona.nombre_titulos3y4
                    data['configFirma'] = configFirma
                    f.editar(configFirma.administrativo.persona_id,
                             configFirma.administrativo.persona.nombre_completo_inverso().upper())
                    f.fields['cargo'].initial = configFirma.cargo
                    data['form'] = f
                    return render(request, "adm_admision/addfirma.html", data)
                except Exception as ex:
                    return render(request, "adm_admision/configuracionfirma.html", data)


            elif action == 'delfirma':
                try:
                    configFirma = ConfigurarFirmaAdmisionPosgrado.objects.get(pk=int(encrypt(request.GET['id'])))
                    configFirma.status = False
                    configFirma.save()
                    log(u'Elimino configuración de firma: %s' % configFirma, request, "del")
                    return JsonResponse({"result": "ok"})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar."})


            elif action == 'buscarpersona':
                try:
                    q = request.GET['q'].upper().strip()
                    s = q.split(" ")
                    cf = ConfigurarFirmaAdmisionPosgrado.objects.values_list('administrativo_id').filter(status=True)
                    per = Administrativo.objects.filter(status=True, persona__status=True).exclude(pk__in=cf)
                    if len(s) == 1:
                        per = Administrativo.objects.filter((
                                Q(persona__nombres__icontains=q) | Q(persona__apellido1__icontains=q) |
                                Q(persona__cedula__icontains=q) | Q(persona__apellido2__icontains=q) |
                                Q(persona__cedula__contains=q))
                        ).filter(status=True, persona__status=True).exclude(pk__in=cf).distinct()[:10]
                    elif len(s) == 2:
                        per = Administrativo.objects.filter(
                            (Q(persona__apellido1__icontains=s[0]) & Q(persona__apellido2__icontains=s[1])) |
                            (Q(persona__nombres__icontains=s[0]) & Q(persona__nombres__icontains=s[1])) |
                            (Q(persona__nombres__icontains=s[0]) & Q(persona__apellido1__contains=s[1]))
                        ).filter(status=True, persona__status=True).exclude(pk__in=cf).distinct()[:10]
                    else:
                        per = Administrativo.objects.filter((
                                                                    Q(persona__nombres__contains=s[0]) & Q(
                                                                persona__apellido1__contains=s[1]) &
                                                                    Q(persona__apellido2__contains=s[2])) | (
                                                                        Q(persona__nombres__contains=s[0]) &
                                                                        Q(persona__nombres__contains=s[1]) & Q(
                                                                    persona__apellido1__contains=s[2]))
                                                            ).filter(status=True, persona__status=True).exclude(
                            pk__in=cf).distinct()[:10]

                    data = {"result": "ok",
                            "results": [{"id": x.persona.id, "name": "{}".format(x.persona.nombre_completo()),
                                         "cargo": x.persona.cargo_persona_2().denominacionpuesto.descripcion if x.persona.cargo_persona_2() else ''}
                                        for x in per]}
                    return JsonResponse(data)
                except Exception as ex:
                    pass


            elif action == 'verdetalle':
                try:
                    data = {}
                    data['cohorte'] = CohorteMaestria.objects.get(pk=int(encrypt(request.GET['id'])))
                    template = get_template("adm_admision/detallecohorte.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            elif action == 'listacomprobantes':
                try:
                    data['title'] = u'Listado de comprobantes'
                    data['listacomprobantes'] = comprobantes = ComprobanteAlumno.objects.filter(status=True).order_by('estados','-fechapago')
                    cursor = connections['epunemi'].cursor()
                    if comprobantes.filter(estados=1):
                        for listacom in comprobantes.filter(estados__in=[1, 3]):
                            sql = """
                                                        SELECT estados, cuenta.numero,tipo.nombre,banco.nombre,
                                                        (SELECT deta.fecha_creacion::TIMESTAMP::DATE FROM sagest_detallecomprobantealumno deta
                                                        WHERE deta.comprobantes_id=compro.id
                                                        ) AS fechavalidada
                                                        FROM sagest_comprobantealumno compro,
                                                        sagest_cuentabanco cuenta,
                                                        sagest_tipocuentabanco tipo,
                                                        sagest_banco banco
                                                        WHERE compro.cuentadeposito_id=cuenta.id
                                                        AND cuenta.tipocuenta_id=tipo.id
                                                        AND cuenta.banco_id=banco.id
                                                        and compro.id = %s
                                                    """ % (listacom.idcomprobanteepunemi)
                            cursor.execute(sql)
                            row = cursor.fetchone()
                            listacom.estados = row[0]
                            listacom.fechavalidacion = row[4]
                            listacom.cuentadeposito = str(row[3]) + ' - #:' + str(row[1]) + ' - Cta:' + str(row[2])
                            listacom.save()
                    return render(request, "adm_admision/listacomprobantes.html", data)
                except Exception as ex:
                    pass

            elif action == 'listarequisitoshomologacion':
                try:
                    id = request.GET.get('id', '0')
                    eMaestria = MaestriasAdmision.objects.get(status=True, pk=int(id))
                    data['title'] = "Listado de requisitos de homologación"

                    eRequisitos = RequisitosMaestria.objects.filter(status=True, maestria=eMaestria).order_by('requisito__nombre')
                    paging = MiPaginador(eRequisitos, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['page'] = page
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['eRequisitos'] = page.object_list
                    data['count'] = eRequisitos.count()
                    data['eMaestria'] = eMaestria
                    return render(request, "adm_admision/listarequisitoshomologacion.html", data)
                except Exception as ex:
                    pass

            return HttpResponseRedirect(request.path)
        else:
            try:
                data['title'] = u'Listado de programas de maestrías'
                s = request.GET.get('s', '')
                ide = request.GET.get('ide', '0')
                filtros = Q(status=True)
                url_vars = ''
                periodosadmision = None

                if puede_ver_todoadmision(request, 'posgrado.change_requisitosmaestria'):
                    if s:
                        filtros = filtros & (Q(descripcion__icontains=s) | Q(carrera__nombre__icontains=s))
                        data['search'] = f"{s}"
                        url_vars += f"&s={s}"

                    if int(ide):
                        if int(ide) == 1:
                            query2 = MaestriasAdmision.objects.filter(filtros).order_by('-id')
                            li = []
                            for que in query2:
                                if que.tiene_requisitos_homologacion():
                                  li.append(que.id)
                            filtros = filtros & (Q(id__in=li))
                            data['ide'] = int(ide)
                            url_vars += f"&ide={ide}"
                        elif int(ide) == 2:
                            query2 = MaestriasAdmision.objects.filter(filtros).order_by('-id')
                            li = []
                            for que in query2:
                                if not que.tiene_requisitos_homologacion():
                                  li.append(que.id)
                            filtros = filtros & (Q(id__in=li))
                            data['ide'] = int(ide)
                            url_vars += f"&ide={ide}"
                        elif int(ide) == 3:
                            query2 = MaestriasAdmision.objects.filter(filtros).order_by('-id')
                            li = []
                            for que in query2:
                                if que.ofertada():
                                  li.append(que.id)
                            filtros = filtros & (Q(id__in=li))
                            data['ide'] = int(ide)
                            url_vars += f"&ide={ide}"
                        elif int(ide) == 4:
                            query2 = MaestriasAdmision.objects.filter(filtros).order_by('-id')
                            li = []
                            for que in query2:
                                if not que.ofertada():
                                  li.append(que.id)
                            filtros = filtros & (Q(id__in=li))
                            data['ide'] = int(ide)
                            url_vars += f"&ide={ide}"

                    periodosadmision = MaestriasAdmision.objects.filter(filtros).distinct().order_by('-id')
                else:
                    if puede_ver_todoadmision(request, 'sga.puede_ver_todo_admisionposgrado'):
                        if s:
                            filtros = filtros & (Q(descripcion__icontains=s) | Q(carrera__nombre__icontains=s))
                            data['search'] = f"{s}"
                            url_vars += f"&s={s}"

                        if int(ide):
                            if int(ide) == 1:
                                query2 = MaestriasAdmision.objects.filter(filtros).order_by('-id')
                                li = []
                                for que in query2:
                                    if que.tiene_requisitos_homologacion():
                                        li.append(que.id)
                                filtros = filtros & (Q(id__in=li))
                                data['ide'] = int(ide)
                                url_vars += f"&ide={ide}"
                            elif int(ide) == 2:
                                query2 = MaestriasAdmision.objects.filter(filtros).order_by('-id')
                                li = []
                                for que in query2:
                                    if not que.tiene_requisitos_homologacion():
                                        li.append(que.id)
                                filtros = filtros & (Q(id__in=li))
                                data['ide'] = int(ide)
                                url_vars += f"&ide={ide}"
                            elif int(ide) == 3:
                                query2 = MaestriasAdmision.objects.filter(filtros).order_by('-id')
                                li = []
                                for que in query2:
                                    if que.ofertada():
                                        li.append(que.id)
                                filtros = filtros & (Q(id__in=li))
                                data['ide'] = int(ide)
                                url_vars += f"&ide={ide}"
                            elif int(ide) == 4:
                                query2 = MaestriasAdmision.objects.filter(filtros).order_by('-id')
                                li = []
                                for que in query2:
                                    if not que.ofertada():
                                        li.append(que.id)
                                filtros = filtros & (Q(id__in=li))
                                data['ide'] = int(ide)
                                url_vars += f"&ide={ide}"

                        periodosadmision = MaestriasAdmision.objects.filter(filtros).distinct().order_by('-id')
                    else:
                        if s:
                            filtros = filtros & (Q(descripcion__icontains=s) | Q(carrera__nombre__icontains=s) & Q(cohortemaestria__coordinador=persona))
                            data['search'] = f"{s}"
                            url_vars += f"&s={s}"

                        if int(ide):
                            if int(ide) == 1:
                                query2 = MaestriasAdmision.objects.filter(filtros).order_by('-id')
                                li = []
                                for que in query2:
                                    if que.tiene_requisitos_homologacion():
                                        li.append(que.id)
                                filtros = filtros & (Q(id__in=li))
                                data['ide'] = int(ide)
                                url_vars += f"&ide={ide}"
                            elif int(ide) == 2:
                                query2 = MaestriasAdmision.objects.filter(filtros).order_by('-id')
                                li = []
                                for que in query2:
                                    if not que.tiene_requisitos_homologacion():
                                        li.append(que.id)
                                filtros = filtros & (Q(id__in=li))
                                data['ide'] = int(ide)
                                url_vars += f"&ide={ide}"
                            elif int(ide) == 3:
                                query2 = MaestriasAdmision.objects.filter(filtros).order_by('-id')
                                li = []
                                for que in query2:
                                    if que.ofertada():
                                        li.append(que.id)
                                filtros = filtros & (Q(id__in=li))
                                data['ide'] = int(ide)
                                url_vars += f"&ide={ide}"
                            elif int(ide) == 4:
                                query2 = MaestriasAdmision.objects.filter(filtros).order_by('-id')
                                li = []
                                for que in query2:
                                    if not que.ofertada():
                                        li.append(que.id)
                                filtros = filtros & (Q(id__in=li))
                                data['ide'] = int(ide)
                                url_vars += f"&ide={ide}"

                        periodosadmision = MaestriasAdmision.objects.filter(filtros & Q(cohortemaestria__coordinador=persona)).distinct().order_by('-id')

                paging = MiPaginador(periodosadmision, 25)
                p = 1
                try:
                    paginasesion = 1
                    if 'paginador' in request.session:
                        paginasesion = int(request.session['paginador'])
                    if 'page' in request.GET:
                        p = int(request.GET['page'])
                    else:
                        p = paginasesion
                    try:
                        page = paging.page(p)
                    except:
                        p = 1
                    page = paging.page(p)
                except:
                    page = paging.page(p)
                request.session['paginador'] = p
                data['paging'] = paging
                data['page'] = page
                data['rangospaging'] = paging.rangos_paginado(p)
                data['periodosadmision'] = page.object_list
                data['url_vars'] = url_vars
                data['eTotal'] = periodosadmision.count() if periodosadmision != None else 0
                return render(request, "adm_admision/view.html", data)
            except Exception as ex:
                pass
