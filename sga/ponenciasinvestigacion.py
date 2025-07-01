# -*- coding: UTF-8 -*-
import json
import os
import random
import zipfile
import collections
import time as ET

import openpyxl
from django.core.files.base import ContentFile
from django.db.models.functions import ExtractYear
from django.template.loader import get_template
import xlwt
from xlwt import *
from django.core.paginator import Paginator
from django.contrib.auth.decorators import login_required
from django.db import transaction, connection, connections
from django.db.models.query_utils import Q
from django.http import HttpResponseRedirect, HttpResponse, JsonResponse
from django.shortcuts import render
from django.template import Context
from decorators import secure_module, last_access
from investigacion.funciones import coordinador_investigacion, notificar_revision_solicitud_produccion_cientifica, codigo_publicacion_valido
from sagest.commonviews import obtener_estados_solicitud, obtener_estado_solicitud
from sagest.models import SolicitudPublicacion, ParticipanteSolicitudPublicacion
from settings import SITE_STORAGE
from sga.commonviews import adduserdata, obtener_reporte
from sga.forms import EvidenciaForm, PonenciaInvestigacionForm, ParticipanteProfesorPonenciaForm, \
    ParticipanteAdministrativoPonenciaForm, PonenciaInvestigacionAuxForm, PlanificarPonenciasForm, \
    PonenciaInvestigacionRegistroForm, ParticipanteInscripcionPonenciaForm
from sga.funciones import MiPaginador, log, generar_nombre, cuenta_email_disponible, email_valido, remover_caracteres, \
    elimina_tildes, remover_caracteres_tildes_unicode, remover_caracteres_especiales_unicode, cuenta_email_disponible_para_envio, validar_archivo
from sga.funcionesxhtml2pdf import conviert_html_to_pdf, convert_html_to_pdf
from sga.models import Evidencia, DetalleEvidencias, PonenciasInvestigacion, TIPO_PARTICIPANTE, ParticipantePonencias, \
    TIPO_PARTICIPANTE_INSTITUCION, CUENTAS_CORREOS, PlanificarPonencias, PlanificarPonenciasRecorrido, miinstitucion, \
    PonenciaComite, Persona
from sga.tasks import send_html_mail, conectar_cuenta
from datetime import datetime,timedelta

from sga.templatetags.sga_extras import encrypt


@login_required(redirect_field_name='ret', login_url='/loginsga')
@secure_module
@last_access
@transaction.atomic()
def view(request):
    data = {}
    adduserdata(request, data)
    persona = request.session['persona']
    miscarreras = persona.mis_carreras()
    if request.method == 'POST':
        action = request.POST['action']

        if action == 'add':
            try:
                f = PonenciaInvestigacionForm(request.POST, request.FILES)

                if 'archivocongreso' in request.FILES:
                    archivo = request.FILES['archivocongreso']
                    descripcionarchivo = 'Publicación'
                    resp = validar_archivo(descripcionarchivo, archivo, ['pdf'], '4MB')
                    if resp['estado'] != "OK":
                        return JsonResponse({"result": "bad", "titulo": "Atención!!!", "mensaje": resp["mensaje"], "showSwal": "True", "swalType": "warning"})

                if 'archivocarta' in request.FILES:
                    archivo = request.FILES['archivocarta']
                    descripcionarchivo = 'Carta de aceptación o invitación'
                    resp = validar_archivo(descripcionarchivo, archivo, ['pdf'], '10MB')
                    if resp['estado'] != "OK":
                        return JsonResponse({"result": "bad", "titulo": "Atención!!!", "mensaje": resp["mensaje"], "showSwal": "True", "swalType": "warning"})

                if 'archivoparticipacion' in request.FILES:
                    archivo = request.FILES['archivoparticipacion']
                    descripcionarchivo = 'Certificado participación'
                    resp = validar_archivo(descripcionarchivo, archivo, ['pdf'], '4MB')
                    if resp['estado'] != "OK":
                        return JsonResponse({"result": "bad", "titulo": "Atención!!!", "mensaje": resp["mensaje"], "showSwal": "True", "swalType": "warning"})

                if 'archivocomite' in request.FILES:
                    archivo = request.FILES['archivocomite']
                    descripcionarchivo = 'Comité científico'
                    resp = validar_archivo(descripcionarchivo, archivo, ['pdf'], '4MB')
                    if resp['estado'] != "OK":
                        return JsonResponse({"result": "bad", "titulo": "Atención!!!", "mensaje": resp["mensaje"], "showSwal": "True", "swalType": "warning"})

                if 'archivoprograma' in request.FILES:
                    archivo = request.FILES['archivoprograma']
                    descripcionarchivo = 'Programa del evento'
                    resp = validar_archivo(descripcionarchivo, archivo, ['pdf'], '4MB')
                    if resp['estado'] != "OK":
                        return JsonResponse({"result": "bad", "titulo": "Atención!!!", "mensaje": resp["mensaje"], "showSwal": "True", "swalType": "warning"})

                if f.is_valid():
                    # Verifico si está repetida la publicación de ese tipo con el mismo título
                    if PonenciasInvestigacion.objects.filter(status=True, nombre=f.cleaned_data['titulo']).exists():
                        return JsonResponse({"result": "bad", "titulo": "Atención!!!", "mensaje": u"El título para el tipo de publicación ya ha sido ingresado anteriormente", "showSwal": "True", "swalType": "warning"})

                    # Verifico si está repetida la solicitud de ese tipo con el mismo resumen
                    if PonenciasInvestigacion.objects.filter(status=True, resumen=f.cleaned_data['resumen']).exists():
                        return JsonResponse({"result": "bad", "titulo": "Atención!!!", "mensaje": u"El resumen para el tipo de publicación ya ha sido ingresado anteriormente", "showSwal": "True", "swalType": "warning"})

                    if f.cleaned_data['fechafin'] < f.cleaned_data['fechainicio']:
                        return JsonResponse({"result": "bad", "titulo": "Atención!!!", "mensaje": u"La fecha de fin debe ser mayor o igual a la fecha de inicio", "showSwal": "True", "swalType": "warning"})

                    # Obtengo los valores del detalle de integrantes del comité
                    integrantes = request.POST.getlist('nombre_integrante[]')  # Todas los nombres
                    instituciones = request.POST.getlist('institucion_integrante[]')  # Todas las instituciones
                    emails = request.POST.getlist('email_integrante[]')  # Todas las direcciones de e-mail

                    # Valido que se hayan ingresado mínimo 3 integrantes del comité
                    if len(integrantes) < 3:
                        return JsonResponse({"result": "bad", "titulo": "Atención!!!", "mensaje": u"El comité científico debe estar conformado al menos de 3 integrantes", "showSwal": "True", "swalType": "warning"})

                    # Valido que no estén repetidos los integrantes del comité
                    listado = [nombre.strip().upper() for nombre in integrantes]
                    repetido = [x for x, y in collections.Counter(listado).items() if y > 1]
                    if repetido:
                        return JsonResponse({"result": "bad", "titulo": "Atención!!!", "mensaje": u"El integrante <b>%s</b> del comité científico está repetido" % (listado[0]), "showSwal": "True", "swalType": "warning"})

                    # Guardo los datos de la publicación
                    ponenciainvestigacion = PonenciasInvestigacion(
                        nombre=f.cleaned_data['titulo'],
                        resumen=f.cleaned_data['resumen'],
                        evento=f.cleaned_data['congreso'],
                        pais=f.cleaned_data['pais'],
                        ciudad=f.cleaned_data['ciudad'],
                        fechainicio=f.cleaned_data['fechainicio'],
                        fechafin=f.cleaned_data['fechafin'],
                        enlace=f.cleaned_data['enlace'],
                        fechapublicacion=f.cleaned_data['fechapublicacion'],
                        estado=1,
                        areaconocimiento=f.cleaned_data['campoamplio'],
                        subareaconocimiento=f.cleaned_data['campoespecifico'],
                        subareaespecificaconocimiento=f.cleaned_data['campodetallado'],
                        lineainvestigacion=f.cleaned_data['lineainvestigacion'],
                        sublineainvestigacion=f.cleaned_data['sublineainvestigacion'],
                        provieneproyecto=f.cleaned_data['provieneproyecto'],
                        tipoproyecto=f.cleaned_data['tipoproyecto'] if f.cleaned_data['provieneproyecto'] else None,
                        pertenecegrupoinv=f.cleaned_data['pertenecegrupoinv'],
                        grupoinvestigacion=f.cleaned_data['grupoinvestigacion'],
                        comitecientifico=True,
                        numerocongreso=f.cleaned_data['numerocongreso'],
                        nombrecomite=f.cleaned_data['nombrecomite'],
                        organizadorevento=f.cleaned_data['organizadorevento'],
                        comiteorganizador=f.cleaned_data['comiteorganizador'],
                        accesoabierto=f.cleaned_data['accesoabierto']
                    )
                    ponenciainvestigacion.save(request)

                    if f.cleaned_data['provieneproyecto']:
                        if int(f.cleaned_data['tipoproyecto']) != 3:
                            ponenciainvestigacion.proyectointerno = f.cleaned_data['proyectointerno']
                        else:
                            ponenciainvestigacion.proyectoexterno = f.cleaned_data['proyectoexterno']

                    ponenciainvestigacion.save(request)

                    # Guardo a los integrantes del comité científico
                    for integrante, institucion, email in zip(integrantes, instituciones, emails):
                        ponenciacomite = PonenciaComite(
                            ponencia=ponenciainvestigacion,
                            integrante=integrante.upper().strip(),
                            institucion=institucion.upper().strip(),
                            email=email.lower().strip()
                        )
                        ponenciacomite.save(request)

                    # Guardo las evidencias
                    # PONENCIA PUBLICADA
                    evidencia = Evidencia.objects.get(pk=6)
                    archivocongreso = request.FILES['archivocongreso']
                    archivocongreso._name = generar_nombre("memoriacongreso", archivocongreso._name)

                    detalleevidencia = DetalleEvidencias(evidencia=evidencia,
                                                         ponencia=ponenciainvestigacion,
                                                         descripcion='PONENCIA PUBLICADA',
                                                         archivo=archivocongreso)
                    detalleevidencia.save(request)

                    # CARTA DE ACEPTACIÓN
                    evidencia = Evidencia.objects.get(pk=4)
                    archivocarta = request.FILES['archivocarta']
                    archivocarta._name = generar_nombre("cartaaceptacion", archivocarta._name)

                    detalleevidencia = DetalleEvidencias(evidencia=evidencia,
                                                         ponencia=ponenciainvestigacion,
                                                         descripcion='CARTA DE ACEPTACIÓN',
                                                         archivo=archivocarta)
                    detalleevidencia.save(request)

                    # CERTIFICADO DE PARTICIPACIÓN
                    evidencia = Evidencia.objects.get(pk=5)
                    archivoparticipacion = request.FILES['archivoparticipacion']
                    archivoparticipacion._name = generar_nombre("certificadoparticipacion", archivoparticipacion._name)

                    detalleevidencia = DetalleEvidencias(evidencia=evidencia,
                                                         ponencia=ponenciainvestigacion,
                                                         descripcion='CERTIFICADO DE PARTICIPACIÓN',
                                                         archivo=archivoparticipacion)
                    detalleevidencia.save(request)

                    # COMITÉ CIENTÍFICO
                    evidencia = Evidencia.objects.get(pk=15)
                    archivocomite = request.FILES['archivocomite']
                    archivocomite._name = generar_nombre("comitecientifico", archivocomite._name)

                    detalleevidencia = DetalleEvidencias(evidencia=evidencia,
                                                         ponencia=ponenciainvestigacion,
                                                         descripcion='COMITÉ CIENTÍFICO',
                                                         archivo=archivocomite)
                    detalleevidencia.save(request)

                    # PROGRAMA DEL EVENTO
                    evidencia = Evidencia.objects.get(pk=27)
                    archivoprograma = request.FILES['archivoprograma']
                    archivoprograma._name = generar_nombre("programaevento", archivoprograma._name)

                    detalleevidencia = DetalleEvidencias(evidencia=evidencia,
                                                         ponencia=ponenciainvestigacion,
                                                         descripcion='PROGRAMA DEL EVENTO',
                                                         archivo=archivoprograma)
                    detalleevidencia.save(request)

                    log(u'%s adicionó ponencia publicada: %s' % (persona, ponenciainvestigacion), request, "add")
                    return JsonResponse({"result": "ok", "titulo": "Proceso exitoso!!!", "mensaje": u"Registro guardado con éxito", "showSwal": True})
                else:
                    msg = ",".join([k + ': ' + v[0] for k, v in f.errors.items()])
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al guardar los datos. [%s]" % msg, "showSwal": "True", "swalType": "error"})
            except Exception as ex:
                msg = ex.__str__()
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al guardar los datos. [%s]" % msg, "showSwal": "True", "swalType": "error"})

        if action == 'ingresarsolicitudponencia':
            try:
                solicitud = SolicitudPublicacion.objects.get(pk=int(encrypt(request.POST['ids'])))

                f = PonenciaInvestigacionRegistroForm(request.POST)

                if f.is_valid():
                    estadosolicitud = int(request.POST['estadosolicitud'])

                    # Si el estado es VERIFICADO
                    if estadosolicitud == 57:
                        if not PonenciasInvestigacion.objects.filter(nombre=f.cleaned_data['nombre']).exists():

                            integrantes = request.POST.getlist('integrante[]')
                            instituciones = request.POST.getlist('institucion[]')
                            emails = request.POST.getlist('email[]')

                            if f.cleaned_data['comitecientifico'] is True:
                                if not integrantes:
                                    return JsonResponse({"result": "bad", "mensaje": u"Ingrese los datos de los integrantes del comité científico evaluador"})

                                if len(integrantes) < 3:
                                    return JsonResponse({"result": "bad", "mensaje": u"Ingrese mínimo 3 integrantes del comité"})

                            if f.cleaned_data['comitecientifico'] is True and integrantes:
                                nenblanco = [nombre for nombre in integrantes if nombre.strip() == '']
                                ienblanco = [institucion for institucion in instituciones if institucion.strip() == '']
                                emailnovalido = [email for email in emails if email.strip() != '' and not email_valido(email)]
                                if nenblanco or ienblanco:
                                    return JsonResponse({"result": "bad", "mensaje": u"Los datos de los integrantes del comité científico evaluador son obligatorios a excepción del e-mail"})

                                if emailnovalido:
                                    return JsonResponse({"result": "bad", "mensaje": u"El formato de una o varias direcciones de e-mail de los integrantes del comité científico evaluador no es válido"})

                                listado = [nombre.strip() for nombre in integrantes if nombre.strip() != '']
                                repetido = [x for x, y in collections.Counter(listado).items() if y > 1]
                                if repetido:
                                    return JsonResponse({"result": "bad", "mensaje": u"El integrante [%s] del comité científico evaluador está repetido" % (repetido[0])})

                            ponencia = PonenciasInvestigacion(
                                                              nombre=f.cleaned_data['nombre'],
                                                              resumen=f.cleaned_data['resumen'],
                                                              evento=f.cleaned_data['evento'],
                                                              pais=f.cleaned_data['pais'],
                                                              ciudad=f.cleaned_data['ciudad'],
                                                              fechainicio=f.cleaned_data['fechainicio'],
                                                              fechafin=f.cleaned_data['fechafin'],
                                                              enlace=f.cleaned_data['enlace'],
                                                              estado=f.cleaned_data['estadopublicacion'],
                                                              fechapublicacion=f.cleaned_data['fechapublicacion'] if int(f.cleaned_data['estadopublicacion']) == 1 else None,
                                                              areaconocimiento=f.cleaned_data['areaconocimiento'],
                                                              subareaconocimiento=f.cleaned_data['subareaconocimiento'],
                                                              subareaespecificaconocimiento=f.cleaned_data['subareaespecificaconocimiento'],
                                                              lineainvestigacion=f.cleaned_data['lineainvestigacion'],
                                                              sublineainvestigacion=f.cleaned_data['sublineainvestigacion'],
                                                              provieneproyecto=f.cleaned_data['provieneproyecto'],
                                                              tipoproyecto=f.cleaned_data['tipoproyecto'] if f.cleaned_data['provieneproyecto'] else None,
                                                              pertenecegrupoinv=f.cleaned_data['pertenecegrupoinv'],
                                                              grupoinvestigacion=f.cleaned_data['grupoinvestigacion'],
                                                              comitecientifico=f.cleaned_data['comitecientifico'],
                                                              numerocongreso=f.cleaned_data['numerocongreso'],
                                                              organizadorevento=f.cleaned_data['organizadorevento'],
                                                              comiteorganizador=f.cleaned_data['comiteorganizador'],
                                                              accesoabierto=f.cleaned_data['accesoabierto'],
                                                              solicitudpublicacion=solicitud
                                                              )
                            ponencia.save(request)

                            if f.cleaned_data['provieneproyecto']:
                                if int(f.cleaned_data['tipoproyecto']) != 3:
                                    ponencia.proyectointerno = f.cleaned_data['proyectointerno']
                                else:
                                    ponencia.proyectoexterno = f.cleaned_data['proyectoexterno']

                            ponencia.save(request)

                            for integrante, institucion, email in zip(integrantes, instituciones, emails):
                                ponenciacomite = PonenciaComite(ponencia=ponencia,
                                                                integrante=integrante,
                                                                institucion=institucion,
                                                                email=email
                                                                )
                                ponenciacomite.save(request)

                            solicitud.registrado = True
                            solicitud.estado_id = estadosolicitud
                            solicitud.save(request)

                            if solicitud.estadopublicacion == 1 and solicitud.archivo:
                                evidencia = Evidencia.objects.get(pk=6)
                                new_file = ContentFile(solicitud.archivo.file.read())
                                # new_file.name = generar_nombre("ponencia_", solicitud.nombre_archivo_publicacion())
                                new_file.name = generar_nombre("memoriacongreso", solicitud.nombre_archivo_publicacion())

                                detalleevidencia = DetalleEvidencias(evidencia=evidencia,
                                                                     ponencia=ponencia,
                                                                     descripcion='PONENCIA PUBLICADA',
                                                                     archivo=new_file)
                                detalleevidencia.save(request)

                            if solicitud.archivocertificado:
                                evidencia = Evidencia.objects.get(pk=4)
                                new_file = ContentFile(solicitud.archivocertificado.file.read())
                                # new_file.name = generar_nombre("ponencia_", solicitud.nombre_archivo_carta_aceptacion())
                                new_file.name = generar_nombre("cartaaceptacion", solicitud.nombre_archivo_carta_aceptacion())

                                detalleevidencia = DetalleEvidencias(evidencia=evidencia,
                                                                     ponencia=ponencia,
                                                                     descripcion='CARTA DE ACEPTACIÓN',
                                                                     archivo=new_file)
                                detalleevidencia.save(request)

                            if solicitud.archivoparticipacion:
                                evidencia = Evidencia.objects.get(pk=5)
                                new_file = ContentFile(solicitud.archivoparticipacion.file.read())
                                # new_file.name = generar_nombre("ponencia_", solicitud.nombre_archivo_certificado_participacion())
                                new_file.name = generar_nombre("certificadoparticipacion", solicitud.nombre_archivo_certificado_participacion())

                                detalleevidencia = DetalleEvidencias(evidencia=evidencia,
                                                                     ponencia=ponencia,
                                                                     descripcion='CERTIFICADO DE PARTICIPACIÓN',
                                                                     archivo=new_file)
                                detalleevidencia.save(request)

                            if solicitud.archivocomite and ponencia.comitecientifico:
                                evidencia = Evidencia.objects.get(pk=15)
                                new_file = ContentFile(solicitud.archivocomite.file.read())
                                new_file.name = generar_nombre("comitecientifico", solicitud.nombre_archivo_comite())

                                detalleevidencia = DetalleEvidencias(evidencia=evidencia,
                                                                     ponencia=ponencia,
                                                                     descripcion='COMITÉ CIENTÍFICO',
                                                                     archivo=new_file)
                                detalleevidencia.save(request)

                            if solicitud.archivoprograma:
                                evidencia = Evidencia.objects.get(pk=27)
                                new_file = ContentFile(solicitud.archivoprograma.file.read())
                                new_file.name = generar_nombre("programaevento", solicitud.nombre_archivo_programa_evento())

                                detalleevidencia = DetalleEvidencias(evidencia=evidencia,
                                                                     ponencia=ponencia,
                                                                     descripcion='PROGRAMA DEL EVENTO',
                                                                     archivo=new_file)
                                detalleevidencia.save(request)

                            # Participantes: Autores/Coautores
                            participantes = ParticipanteSolicitudPublicacion.objects.filter(solicitud=solicitud, status=True).order_by('id')
                            for p in participantes:
                                participanteponencia = ParticipantePonencias(matrizevidencia_id=4,
                                                                              ponencias=ponencia,
                                                                              profesor=p.profesor,
                                                                              tipoparticipante=p.tipo,
                                                                              administrativo=p.administrativo,
                                                                              tipoparticipanteins=p.tipoparticipanteins,
                                                                              inscripcion=p.inscripcion)
                                participanteponencia.save(request)


                            notificar_revision_solicitud_produccion_cientifica(solicitud)
                            log(u'Adicionó ponencia de investigacion: %s' % ponencia, request, "add")
                            return JsonResponse({"result": "ok", "titulo": "Proceso exitoso!!!", "mensaje": u"Registro guardado con éxito", "showSwal": True})
                        else:
                            return JsonResponse({"result": "bad", "titulo": "Atención!!!", "mensaje": u"La ponencia ya ha sido ingresada", "showSwal": "True", "swalType": "warning"})
                    else:
                        observacion = request.POST['observacionsolicitud'].strip().upper()

                        # Actualiza solicitud
                        solicitud.estado_id = estadosolicitud
                        solicitud.observacion = observacion
                        solicitud.save(request)

                        notificar_revision_solicitud_produccion_cientifica(solicitud)

                        # Si estado es NOVEDADES
                        if estadosolicitud == 58:
                            log(u'Registró novedad en solicitud: %s' % solicitud.persona, request, "edit")
                        else:  # RECHAZADO
                            log(u'Rechazó solicitud de ponencia: %s' % solicitud.persona, request, "edit")

                        return JsonResponse({"result": "ok", "titulo": "Proceso exitoso!!!", "mensaje": u"Registro guardado con éxito", "showSwal": True})
                else:
                     raise NameError('Error')
            except Exception as ex:
                msg = ex.__str__()
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al guardar los datos. [%s]" % msg, "showSwal": "True", "swalType": "error"})

        elif action == 'editarponencia':
            try:
                ponencia = PonenciasInvestigacion.objects.get(pk=int(encrypt(request.POST['id'])))
                f = PonenciaInvestigacionRegistroForm(request.POST)

                if f.is_valid():
                    if not PonenciasInvestigacion.objects.filter(nombre=f.cleaned_data['nombre']).exclude(pk=int(encrypt(request.POST['id']))).exists():

                        integrantes = request.POST.getlist('integrante[]')
                        instituciones = request.POST.getlist('institucion[]')
                        emails = request.POST.getlist('email[]')

                        if f.cleaned_data['comitecientifico'] is True:
                            if not integrantes:
                                return JsonResponse({"result": "bad", "mensaje": u"Ingrese los datos de los integrantes del comité científico evaluador"})

                            if len(integrantes) < 3:
                                return JsonResponse({"result": "bad", "mensaje": u"Ingrese mínimo 3 integrantes del comité"})

                        if f.cleaned_data['comitecientifico'] is True and integrantes:
                            nenblanco = [nombre for nombre in integrantes if nombre.strip() == '']
                            ienblanco = [institucion for institucion in instituciones if institucion.strip() == '']
                            emailnovalido = [email for email in emails if email.strip() != '' and not email_valido(email)]
                            if nenblanco or ienblanco:
                                return JsonResponse({"result": "bad", "mensaje": u"Los datos de los integrantes del comité científico evaluador son obligatorios a excepción del e-mail"})

                            if emailnovalido:
                                return JsonResponse({"result": "bad", "mensaje": u"El formato de una o varias direcciones de e-mail de los integrantes del comité científico evaluador no es válido"})

                            listado = [nombre.strip() for nombre in integrantes if nombre.strip() != '']
                            repetido = [x for x, y in collections.Counter(listado).items() if y > 1]
                            if repetido:
                                return JsonResponse({"result": "bad", "mensaje": u"El integrante [%s] del comité científico evaluador está repetido" % (repetido[0])})

                        ponencia.nombre = f.cleaned_data['nombre']
                        ponencia.resumen = f.cleaned_data['resumen']
                        ponencia.evento = f.cleaned_data['evento']
                        ponencia.pais = f.cleaned_data['pais']
                        ponencia.ciudad = f.cleaned_data['ciudad']
                        ponencia.fechainicio = f.cleaned_data['fechainicio']
                        ponencia.fechafin = f.cleaned_data['fechafin']
                        ponencia.enlace = f.cleaned_data['enlace']
                        ponencia.estado = f.cleaned_data['estadopublicacion']
                        ponencia.fechapublicacion = f.cleaned_data['fechapublicacion'] if int(f.cleaned_data['estadopublicacion']) == 1 else None
                        ponencia.areaconocimiento = f.cleaned_data['areaconocimiento']
                        ponencia.subareaconocimiento = f.cleaned_data['subareaconocimiento']
                        ponencia.subareaespecificaconocimiento = f.cleaned_data['subareaespecificaconocimiento']
                        ponencia.lineainvestigacion = f.cleaned_data['lineainvestigacion']
                        ponencia.sublineainvestigacion = f.cleaned_data['sublineainvestigacion']
                        ponencia.provieneproyecto = f.cleaned_data['provieneproyecto']
                        ponencia.tipoproyecto = f.cleaned_data['tipoproyecto'] if f.cleaned_data['provieneproyecto'] else None
                        ponencia.pertenecegrupoinv = f.cleaned_data['pertenecegrupoinv']
                        ponencia.grupoinvestigacion = f.cleaned_data['grupoinvestigacion']
                        ponencia.comitecientifico = f.cleaned_data['comitecientifico']
                        ponencia.numerocongreso = f.cleaned_data['numerocongreso']
                        ponencia.organizadorevento = f.cleaned_data['organizadorevento']
                        ponencia.comiteorganizador = f.cleaned_data['comiteorganizador']
                        ponencia.accesoabierto = f.cleaned_data['accesoabierto']

                        if f.cleaned_data['provieneproyecto']:
                            if int(f.cleaned_data['tipoproyecto']) != 3:
                                ponencia.proyectointerno = f.cleaned_data['proyectointerno']
                            else:
                                ponencia.proyectoexterno = f.cleaned_data['proyectoexterno']
                        else:
                            ponencia.tipoproyecto = None
                            ponencia.proyectoexterno = None
                            ponencia.proyectointerno = None

                        ponencia.save(request)

                        if ponencia.integrantecomite():
                            PonenciaComite.objects.filter(ponencia=ponencia, status=True).update(status=False)

                        if f.cleaned_data['comitecientifico'] is True:
                            for integrante, institucion, email in zip(integrantes, instituciones, emails):
                                ponenciacomite = PonenciaComite(ponencia=ponencia,
                                                                integrante=integrante,
                                                                institucion=institucion,
                                                                email=email
                                                                )
                                ponenciacomite.save(request)

                        log(u'Editó ponencia de investigacion: %s' % ponencia, request, "edit")
                        return JsonResponse({"result": "ok", "titulo": "Proceso exitoso!!!", "mensaje": u"Registro actualizado con éxito", "showSwal": True})
                    else:
                        return JsonResponse({"result": "bad", "titulo": "Atención!!!", "mensaje": u"La ponencia ya ha sido ingresado", "showSwal": "True", "swalType": "warning"})
            except Exception as ex:
                msg = ex.__str__()
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al guardar los datos. [%s]" % msg, "showSwal": "True", "swalType": "error"})

        elif action == 'deleteparticipanteponencia':
            try:
                participante = ParticipantePonencias.objects.get(pk=request.POST['id'])
                participante.status = False
                participante.save(request)
                log(u'Eliminó participante de ponencia de investigacion: %s' % participante, request, "del")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'deleteponencia':
            try:
                ponencias =PonenciasInvestigacion.objects.get(pk=request.POST['id'])
                ponencias.status = False
                ponencias.save(request)
                log(u'Eliminó ponencia de investigacion: %s' % ponencias, request, "del")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'addparticipantesdocentes':
            try:
                f = ParticipanteProfesorPonenciaForm(request.POST)
                if f.is_valid():
                    if ParticipantePonencias.objects.filter(status=True, matrizevidencia_id=4, ponencias_id=int(encrypt(request.POST['id'])), profesor_id=f.cleaned_data['profesor']).exists():
                        return JsonResponse({"result": "bad", "mensaje": u"El participante ya fue agregado anteriormente."})
                    else:
                        programas = ParticipantePonencias(matrizevidencia_id=4,
                                                          ponencias_id=int(encrypt(request.POST['id'])),
                                                          profesor_id=f.cleaned_data['profesor'],
                                                          tipoparticipante=f.cleaned_data['tipoparticipante'],
                                                          tipoparticipanteins=f.cleaned_data['tipoparticipanteins']
                                                          )
                        programas.save(request)
                        log(u'Adicionó participante docente a ponencia de investigacion: %s' % programas, request, "add")
                        return JsonResponse({"result": "ok"})
                else:
                     raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'addparticipantesadministrativos':
            try:
                f = ParticipanteAdministrativoPonenciaForm(request.POST)
                if f.is_valid():
                    if ParticipantePonencias.objects.filter(status=True, matrizevidencia_id=4, ponencias_id=int(encrypt(request.POST['id'])), administrativo_id=f.cleaned_data['administrativo']).exists():
                        return JsonResponse({"result": "bad", "mensaje": u"El participante ya fue agregado anteriormente."})
                    else:
                        programas = ParticipantePonencias(matrizevidencia_id=4,
                                                          ponencias_id=int(encrypt(request.POST['id'])),
                                                          administrativo_id=f.cleaned_data['administrativo'],
                                                          tipoparticipante=f.cleaned_data['tipoparticipante'],
                                                          tipoparticipanteins=f.cleaned_data['tipoparticipanteins']
                                                          )
                        programas.save(request)
                        log(u'Adicionó participante administrativo a ponencia de investigacion: %s' % programas, request, "add")
                        return JsonResponse({"result": "ok"})
                else:
                     raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'addparticipantesinscripcion':
            try:
                f = ParticipanteInscripcionPonenciaForm(request.POST)
                if f.is_valid():
                    if ParticipantePonencias.objects.filter(status=True, matrizevidencia_id=4, ponencias_id=int(encrypt(request.POST['id'])), inscripcion_id=f.cleaned_data['inscripcion']).exists():
                        return JsonResponse({"result": "bad", "mensaje": u"El participante ya fue agregado anteriormente."})
                    else:

                        programas = ParticipantePonencias(matrizevidencia_id=4,
                                                          ponencias_id=int(encrypt(request.POST['id'])),
                                                          inscripcion_id=f.cleaned_data['inscripcion'],
                                                          tipoparticipante=f.cleaned_data['tipoparticipante'],
                                                          tipoparticipanteins=f.cleaned_data['tipoparticipanteins']
                                                          )
                        programas.save(request)
                        log(u'Adiciono participantes estudiante en ponencias: %s [%s]' % (programas, programas.id), request, "add")
                        return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'addevidenciasponencia':
            try:
                f = EvidenciaForm(request.POST, request.FILES)
                # 2.5MB - 2621440
                # 5MB - 5242880
                # 10MB - 10485760
                # 20MB - 20971520
                # 50MB - 5242880
                # 100MB 104857600
                # 250MB - 214958080
                # 500MB - 429916160
                d = request.FILES['archivo']
                if d.size > 10485760:
                    return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 10 Mb."})
                if f.is_valid():
                    newfile = None
                    if 'archivo' in request.FILES:
                        newfile = request.FILES['archivo']
                        newfile._name = generar_nombre("ponencia_", newfile._name)
                    if DetalleEvidencias.objects.filter(evidencia_id=int(encrypt(request.POST['idevidencia'])), ponencia_id=int(encrypt(request.POST['id']))).exists():
                        detalle = DetalleEvidencias.objects.get(evidencia_id=int(encrypt(request.POST['idevidencia'])), ponencia_id=int(encrypt(request.POST['id'])))
                        detalle.descripcion = f.cleaned_data['descripcion']
                        detalle.archivo = newfile
                        detalle.save(request)
                        log(u'Adicionó detalle de evidencia de ponencia de investigacion: %s' % detalle, request,"add")
                    else:
                        evidencia = DetalleEvidencias(evidencia_id=int(encrypt(request.POST['idevidencia'])),
                                                      ponencia_id=int(encrypt(request.POST['id'])),
                                                      descripcion=f.cleaned_data['descripcion'],
                                                      archivo=newfile)
                        evidencia.save(request)
                        log(u'Adicionó detalle de evidencia de ponencia de investigacion: %s' % evidencia,request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                     raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'solicitudes':
            try:
                data['solicitudes'] = SolicitudPublicacion.objects.filter(tiposolicitud=int(request.POST['tipo']), status=True, estado__valor=1).order_by('fecha_creacion')
                template = get_template("inv_ponencia/solicitudes.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'html': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'aprobarponencia':
            try:
                ponencia = PonenciasInvestigacion.objects.get(pk=int(encrypt(request.POST['id'])))

                # Si existe solicitud
                if ponencia.solicitudpublicacion:
                    estadosolicitud = obtener_estado_solicitud(8, 5)  # APROBADO

                    solicitud = SolicitudPublicacion.objects.get(pk=ponencia.solicitudpublicacion.id)
                    solicitud.aprobado = True
                    solicitud.estado = estadosolicitud
                    solicitud.save(request)

                ponencia.aprobado = True
                ponencia.save(request)

                integrantes = ponencia.participantes()
                ponencia.tipoaporte = 1 if integrantes.filter(tipoparticipanteins=1).count() >= 1 else 2
                ponencia.save(request)

                # Si existe solicitud se debe notificar
                if ponencia.solicitudpublicacion:
                    notificar_revision_solicitud_produccion_cientifica(solicitud)

                log(u'%s Aprobó ponencia: %s' % (persona, ponencia.nombre), request, "edit")
                return JsonResponse({"result": "ok", "titulo": "Proceso exitoso!!!", "mensaje": u"Registro guardado con éxito", "showSwal": True})
            except Exception as ex:
                msg = ex.__str__()
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al guardar los datos. [%s]" % msg, "showSwal": "True", "swalType": "error"})

        elif action == 'updatetipoparticipante':
            try:
                participantes = ParticipantePonencias.objects.get(pk=request.POST['iditem'])
                participantes.tipoparticipanteins = request.POST['idtipo']
                participantes.save(request)
                return JsonResponse({'result': 'ok', 'id': participantes.ponencias.id})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad'})

        elif action == 'detallecongreso':
            try:
                data['congreso'] = congreso = PlanificarPonencias.objects.get(id=(request.POST['id']))

                if congreso.id > 433:
                    data['criteriosponencia'] = congreso.planificarponenciascriterio_set.filter(status=True).order_by('id')

                template = get_template("inv_ponencia/detallecongreso.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'html': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'detallerecorrido':
            try:
                data['aprobador'] = persona
                data['congreso'] = congreso = PlanificarPonencias.objects.get(id=(request.POST['id']))
                data['congresorecorrido'] = congreso.planificarponenciasrecorrido_set.filter(status=True)
                template = get_template("inv_ponencia/detallerecorrido.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'html': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'addplanificarponenciarecorrido':
            try:
                solicitud = PlanificarPonencias.objects.get(pk=request.POST['id'])
                estado = int(request.POST['esta'])
                observacion = request.POST['obse'].strip().upper()

                solicitud.estado = estado
                solicitud.save(request)

                recorrido = PlanificarPonenciasRecorrido(planificarponencias=solicitud,
                                            fecha=datetime.now().date(),
                                            observacion=observacion,
                                            persona=persona,
                                            estado=int(request.POST['esta']))

                recorrido.save(request)

                # Notificar por e-mail
                # Obtengo la(s) cuenta(s) de correo desde la cual(es) se envía el e-mail
                listacuentascorreo = [29]  # investigacion.dip@unemi.edu.ec

                # E-mail del destinatario solicitante
                lista_email_envio = solicitud.profesor.persona.lista_emails_envio()
                # lista_email_envio = ['isaltosm@unemi.edu.ec']
                lista_email_cco = ['ivan_saltos_medina@hotmail.com']
                lista_adjuntos = []

                fechaenvio = datetime.now().date()
                horaenvio = datetime.now().time()
                cuenta = cuenta_email_disponible_para_envio(listacuentascorreo)

                if estado == 2:
                    tituloemail = "Solicitud de Financiamiento a Ponencia Pre-Seleccionada"
                    tiponotificacion = "PRESELSOL"
                    estadoasignado = "PRE-SELECCIONÓ"
                elif estado == 3:
                    tituloemail = "Solicitud de Financiamiento a Ponencia Aprobada"
                    tiponotificacion = "APROSOL"
                    estadoasignado = "APROBÓ"
                elif estado == 4:
                    tituloemail = "Solicitud de Financiamiento a Ponencia Rechazada"
                    tiponotificacion = "RECHSOL"
                    estadoasignado = "RECHAZÓ"
                elif estado == 5:
                    tituloemail = "Solicitud de Financiamiento a Ponencia Autorizada"
                    tiponotificacion = "AUTSOL"
                    estadoasignado = "AUTORIZÓ"
                else:
                    tituloemail = "Novedades con Solicitud de Financiamiento a Ponencia"
                    tiponotificacion = "NOVSOL"
                    estadoasignado = "NOVEDAD"

                titulo = "Postulación y Financiamiento de Ponencias"

                send_html_mail(tituloemail,
                               "emails/financiamientoponencia.html",
                               {'sistema': u'SGA - UNEMI',
                                'titulo': titulo,
                                'fecha': fechaenvio,
                                'hora': horaenvio,
                                'tiponotificacion': tiponotificacion,
                                'saludo': 'Estimada' if solicitud.profesor.persona.sexo_id == 1 else 'Estimado',
                                'nombrepersona': solicitud.profesor.persona.nombre_completo_inverso(),
                                'solicitud': solicitud,
                                'estadoasignado': estadoasignado,
                                'observacion': observacion
                                },
                               lista_email_envio,
                               lista_email_cco,
                               lista_adjuntos,
                               cuenta=CUENTAS_CORREOS[cuenta][1]
                               )

                log(u'Adicionó recorrido en planificación de ponencias: %s - %s - Estado %s' % (recorrido.planificarponencias, recorrido.persona, recorrido.estado), request, "add")
                return JsonResponse({"result": "ok", "titulo": "Proceso exitoso!!!", "mensaje": u"Registro guardado con éxito", "showSwal": True})


                # if congreso.estado == 2:
                #     asunto = u"ESTADO DE APROBACIÓN DE PONENCIA"
                #     send_html_mail(asunto, "emails/preseleccion_ponencias.html",
                #                    {'sistema': request.session['nombresistema'],
                #                     'congreso': congreso,
                #                     'profesor': congreso.profesor.persona.nombre_completo_inverso()},
                #                    #['kpalaciosz@unemi.edu.ec']
                #                     congreso.profesor.persona.lista_emails_envio()
                #                    , [], cuenta=CUENTAS_CORREOS[0][1])
                # if congreso.estado == 3:
                #     asunto = u"ESTADO DE APROBACIÓN DE PONENCIA"
                #     send_html_mail(asunto, "emails/aprobacion_ponencias.html",
                #                    {'sistema': request.session['nombresistema'],
                #                     'congreso': congreso,
                #                     'profesor': congreso.profesor.persona.nombre_completo_inverso()},
                #                    #['kpalaciosz@unemi.edu.ec']
                #                     congreso.profesor.persona.lista_emails_envio()
                #                    , [], cuenta=CUENTAS_CORREOS[0][1])
                #
                # if congreso.estado == 5:
                #     asunto = u"ESTADO DE APROBACIÓN DE PONENCIA"
                #     send_html_mail(asunto, "emails/autorizacion_ponencias.html",
                #                    {'sistema': request.session['nombresistema'],
                #                     'congreso': congreso,
                #                     'profesor': congreso.profesor.persona.nombre_completo_inverso()},
                #                    #['kpalaciosz@unemi.edu.ec']
                #                     congreso.profesor.persona.lista_emails_envio()
                #                    , [], cuenta=CUENTAS_CORREOS[0][1])

                # return JsonResponse({"result": "ok"})
            except Exception as ex:
                msg = ex.__str__()
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al guardar los datos. [%s]" % msg, "showSwal": "True", "swalType": "error"})

        elif action == 'rechazar':
            try:
                solicitud = SolicitudPublicacion.objects.get(pk=request.POST['id'])
                observacion = request.POST['observacion'].strip().upper()
                if observacion:
                    solicitud.aprobado = False
                    solicitud.observacion = observacion
                    solicitud.save(request)

                    log(u'Rechazo solicitud de ponencia: %s' % solicitud.persona, request, "edit")

                    cuenta = cuenta_email_disponible()

                    asunto = u"Solicitud de Registro de Ponencia Rechazada"
                    send_html_mail(asunto, "emails/estadosolicitudponencia.html",
                                   {'sistema': 'SGA - UNEMI',
                                    'saludo': 'Estimado' if solicitud.persona.sexo.id == 2 else 'Estimada',
                                    'titulo': solicitud.nombre,
                                    'motivorechazo': observacion,
                                    'solicitante': solicitud.persona.nombre_completo_inverso()},
                                   solicitud.persona.lista_emails_envio(),
                                   [],
                                   cuenta=CUENTAS_CORREOS[cuenta][1])
                    return JsonResponse({"result": "ok"})
                else:
                    return JsonResponse({"result": "bad", "mensaje": u"Debe ingresar observación."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar el registro."})

        elif action == 'editponencia':
            try:
                form = PlanificarPonenciasForm(request.POST)
                if form.is_valid():
                    tipo = request.POST['tipo']
                    lista = json.loads(request.POST['lista_items1'])
                    criterioi1 = criterioi2 = criterioi3 = criterioi4 = criterion1 = criterion2 = criteriog1 = criteriog2 = False
                    for l in lista:
                        if l['id'] == '1':
                            criterioi1 = l['valor']
                        if l['id'] == '2':
                            criterioi2 = l['valor']

                        if l['id'] == '4':
                            criterioi4 = l['valor']
                        if l['id'] == '7':
                            criterion1 = l['valor']
                        if l['id'] == '8':
                            criterion2 = l['valor']
                        if tipo == 'I':
                            if l['id'] == '5':
                                criteriog1 = l['valor']
                            if l['id'] == '6':
                                criteriog2 = l['valor']
                            if l['id'] == '3':
                                criterioi3 = l['valor']

                        if tipo == 'N':
                            if l['id'] == '9':
                                criteriog1 = l['valor']
                            if l['id'] == '10':
                                criteriog2 = l['valor']
                            if l['id'] == '11':
                                criterioi3 = l['valor']

                    planificarponencias = PlanificarPonencias.objects.get(pk=int(request.POST['id']))
                    planificarponencias.tema = form.cleaned_data['tema']
                    planificarponencias.justificacion = form.cleaned_data['justificacion']
                    planificarponencias.fecha_inicio = form.cleaned_data['fechainicio']
                    planificarponencias.link = form.cleaned_data['link']
                    planificarponencias.fecha_fin = form.cleaned_data['fechafin']
                    planificarponencias.nombre = form.cleaned_data['nombre']
                    planificarponencias.pais = form.cleaned_data['pais']
                    planificarponencias.sugerenciacongreso = form.cleaned_data['sugerenciacongreso'] if 'sugerenciacongreso' in form.cleaned_data else None
                    planificarponencias.criterioi1 = criterioi1
                    planificarponencias.criterioi2 = criterioi2
                    planificarponencias.criterioi3 = criterioi3
                    planificarponencias.criterioi4 = criterioi4
                    planificarponencias.criterion1 = True
                    planificarponencias.criterion2 = criterion2
                    planificarponencias.criteriog1 = criteriog1
                    planificarponencias.criteriog2 = criteriog2
                    planificarponencias.save(request)
                    log(u'Edito planificación de ponencias: %s' % planificarponencias, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'reportefichacatalografica':
            try:
                data = {}
                ponencia = PonenciasInvestigacion.objects.get(pk=int(encrypt(request.POST['id'])))
                participantes = ponencia.participantes()
                data['ponencia'] = ponencia
                data['participantes'] = participantes

                return conviert_html_to_pdf(
                    'inv_ponencia/fichacatalografica_pdf.html',
                    {
                        'pagesize': 'A4',
                        'data': data,
                    }
                )
            except Exception as ex:
                return HttpResponseRedirect("/ponenciasinvestigacion?info=%s" % "Error al generar el reporte de ficha catalográfica")

        elif action == 'evidenciasponencia':
            try:
                # Aquí se almacenan las fichas catalográficas
                directorio = SITE_STORAGE + '/media/articulos'
                try:
                    os.stat(directorio)
                except:
                    os.mkdir(directorio)

                caracteres_a_quitar = "!\"\#$%&()=?¡'¿{}[],.-+*/|°^~:;"
                anio = request.POST['anio']
                output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media', 'zipav'))

                response = HttpResponse(content_type='application/zip')
                response['Content-Disposition'] = 'attachment; filename=ponenciasevidencias_' + random.randint(1, 10000).__str__() + '.zip'

                nombre_archivo = "evidenciasponencias" + anio + ".zip"
                filename = os.path.join(output_folder, nombre_archivo)
                fantasy_zip = zipfile.ZipFile(filename, 'w')

                ponencias_publicadas = PonenciasInvestigacion.objects.filter(status=True, fechapublicacion__year=anio).order_by('id')

                for ponencia in ponencias_publicadas:
                    titulo = ponencia.nombre

                    palabras = titulo.split(" ")
                    titulo = "_".join(palabras[0:5])
                    titulo = remover_caracteres(titulo, caracteres_a_quitar)

                    titulo = elimina_tildes(remover_caracteres_tildes_unicode(remover_caracteres_especiales_unicode(titulo)))
                    carpetaponencia = "PUBLICADOS/PON_" + str(ponencia.id) + "_" + titulo

                    ponencia_id = ponencia.id
                    nombre = "PONENCIA_" + str(ponencia_id).zfill(4)

                    # Agrego las evidencias a la carpeta de la ponencia
                    for evidencia in ponencia.detalleevidencias_set.filter(status=True):
                        ext = evidencia.archivo.__str__()[evidencia.archivo.__str__().rfind("."):]

                        if evidencia.descripcion:
                            nombreevidencia = evidencia.descripcion.upper()
                            nombreevidencia = remover_caracteres(nombreevidencia, caracteres_a_quitar)
                            nombreevidencia = elimina_tildes(remover_caracteres_tildes_unicode(remover_caracteres_especiales_unicode(nombreevidencia))).replace(" ", "_")
                        else:
                            nombreevidencia = "EVIDENCIA" + random.randint(1, 10000).__str__()


                        if evidencia.evidencia.id == 4:
                            if os.path.exists(SITE_STORAGE + evidencia.archivo.url):
                                fantasy_zip.write(SITE_STORAGE + evidencia.archivo.url, carpetaponencia + '/' + nombreevidencia + ext.lower())

                        elif evidencia.evidencia.id == 5:
                            if os.path.exists(SITE_STORAGE + evidencia.archivo.url):
                                fantasy_zip.write(SITE_STORAGE + evidencia.archivo.url, carpetaponencia + '/' + nombreevidencia + ext.lower())

                        elif evidencia.evidencia.id == 6:
                            if os.path.exists(SITE_STORAGE + evidencia.archivo.url):
                                fantasy_zip.write(SITE_STORAGE + evidencia.archivo.url, carpetaponencia + '/' + nombreevidencia + ext.lower())

                        elif evidencia.evidencia.id == 15:
                            if os.path.exists(SITE_STORAGE + evidencia.archivo.url):
                                fantasy_zip.write(SITE_STORAGE + evidencia.archivo.url, carpetaponencia + '/' + nombreevidencia + ext.lower())

                    # Generar la ficha catalográfica de la ponencia
                    data['ponencia'] = ponencia
                    data['participantes'] = ponencia.participantes()

                    nombrearchivoficha = 'fichacatalograficapon_' + str(ponencia_id) + '_' + str(random.randint(1, 10000)) + '.pdf'
                    valida = convert_html_to_pdf(
                        'inv_ponencia/fichacatalografica_pdf.html',
                        {'pagesize': 'A4', 'data': data},
                        nombrearchivoficha,
                        directorio
                    )

                    archivoficha = directorio + "/" + nombrearchivoficha
                    # Agrego el archivo de la ficha a la carpeta del artículo
                    fantasy_zip.write(archivoficha, carpetaponencia + "/FICHA_CATALOGRAFICA.pdf")
                    # Borro el archivo de la ficha creado
                    os.remove(archivoficha)

                ponencias_no_publicadas = PonenciasInvestigacion.objects.filter(status=True, fechapublicacion__isnull=True, fechainicio__year=anio).order_by('id')

                for ponencia in ponencias_no_publicadas:
                    titulo = ponencia.nombre

                    palabras = titulo.split(" ")
                    titulo = "_".join(palabras[0:5])
                    titulo = remover_caracteres(titulo, caracteres_a_quitar)

                    titulo = elimina_tildes(remover_caracteres_tildes_unicode(remover_caracteres_especiales_unicode(titulo)))
                    carpetaponencia = "NOPUBLICADAS/PON_" + str(ponencia.id) + "_" + titulo

                    ponencia_id = ponencia.id
                    nombre = "PONENCIA_" + str(ponencia_id).zfill(4)

                    # Agrego las evidencias a la carpeta de la ponencia
                    for evidencia in ponencia.detalleevidencias_set.filter(status=True):
                        ext = evidencia.archivo.__str__()[evidencia.archivo.__str__().rfind("."):]

                        if evidencia.descripcion:
                            nombreevidencia = evidencia.descripcion.upper()
                            nombreevidencia = remover_caracteres(nombreevidencia, caracteres_a_quitar)
                            nombreevidencia = elimina_tildes(remover_caracteres_tildes_unicode(remover_caracteres_especiales_unicode(nombreevidencia))).replace(" ", "_")
                        else:
                            nombreevidencia = "EVIDENCIA" + random.randint(1, 10000).__str__()

                        if evidencia.evidencia.id == 4:
                            if os.path.exists(SITE_STORAGE + evidencia.archivo.url):
                                fantasy_zip.write(SITE_STORAGE + evidencia.archivo.url, carpetaponencia + '/' + nombreevidencia + ext.lower())

                        elif evidencia.evidencia.id == 5:
                            if os.path.exists(SITE_STORAGE + evidencia.archivo.url):
                                fantasy_zip.write(SITE_STORAGE + evidencia.archivo.url, carpetaponencia + '/' + nombreevidencia + ext.lower())

                        elif evidencia.evidencia.id == 6:
                            if os.path.exists(SITE_STORAGE + evidencia.archivo.url):
                                fantasy_zip.write(SITE_STORAGE + evidencia.archivo.url, carpetaponencia + '/' + nombreevidencia + ext.lower())

                        elif evidencia.evidencia.id == 15:
                            if os.path.exists(SITE_STORAGE + evidencia.archivo.url):
                                fantasy_zip.write(SITE_STORAGE + evidencia.archivo.url, carpetaponencia + '/' + nombreevidencia + ext.lower())

                    # Generar la ficha catalográfica de la ponencia
                    data['ponencia'] = ponencia
                    data['participantes'] = ponencia.participantes()

                    nombrearchivoficha = 'fichacatalograficapon_' + str(ponencia_id) + '_' + str(random.randint(1, 10000)) + '.pdf'
                    valida = convert_html_to_pdf(
                        'inv_ponencia/fichacatalografica_pdf.html',
                        {'pagesize': 'A4', 'data': data},
                        nombrearchivoficha,
                        directorio
                    )

                    archivoficha = directorio + "/" + nombrearchivoficha
                    # Agrego el archivo de la ficha a la carpeta del artículo
                    fantasy_zip.write(archivoficha, carpetaponencia + "/FICHA_CATALOGRAFICA.pdf")
                    # Borro el archivo de la ficha creado
                    os.remove(archivoficha)


                fantasy_zip.close()

                ruta = "media/zipav/" + nombre_archivo

                return JsonResponse({'result': 'ok', 'archivo': ruta})
            except Exception as ex:
                msg = ex.__str__()
                return JsonResponse({"result": "bad", "mensaje": u"Error al descargar evidencias. Detalle: %s" % (msg)})

        elif action == 'evidenciasponenciacodigo':
            try:
                archivo = request.FILES['archivo']
                descripcionarchivo = 'Archivo Excel de Códigos'
                resp = validar_archivo(descripcionarchivo, archivo, ['xlsx'], '4MB')
                if resp['estado'] != "OK":
                    return JsonResponse({"result": "bad", "titulo": "Atención!!!", "mensaje": resp["mensaje"], "showSwal": "True", "swalType": "warning"})

                # Aquí se almacenan las fichas catalográficas
                directorio = SITE_STORAGE + '/media/articulos'
                try:
                    os.stat(directorio)
                except:
                    os.mkdir(directorio)

                caracteres_a_quitar = "!\"\#$%&()=?¡'¿{}[],.-+*/|°^~:;"
                output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media', 'zipav'))

                response = HttpResponse(content_type='application/zip')
                response['Content-Disposition'] = 'attachment; filename=ponenciasevidencias_' + random.randint(1, 10000).__str__() + '.zip'

                nombre_archivo = "evidenciasponenciascodigo.zip"
                filename = os.path.join(output_folder, nombre_archivo)
                fantasy_zip = zipfile.ZipFile(filename, 'w')

                wb = openpyxl.load_workbook(archivo)
                sheet = wb.worksheets[0]
                lista_ids = []
                lista_filas = []
                novalidos = 0

                # Recorrer las filas de la hoja
                for fila in range(1, sheet.max_row + 1):
                    if fila > 1:
                        codigo = sheet.cell(row=fila, column=1).value
                        if codigo_publicacion_valido(codigo, "PON"):
                            codigo = codigo.strip()
                            lista_ids.append(codigo.split("-")[0])
                        else:
                            lista_filas.append(fila)
                            novalidos += 1

                if novalidos > 0:
                    return JsonResponse({"result": "bad", "titulo": "Atención!!!", "mensaje": "{} {} {} que {} el código en formato incorrecto. Filas: {}".format("Existe" if novalidos == 1 else "Existen", novalidos, "registro" if novalidos == 1 else "registros", "tiene" if novalidos == 1 else "tienen", ", ".join(str(f) for f in lista_filas)), "showSwal": "True", "swalType": "warning"})

                anios = PonenciasInvestigacion.objects.filter(status=True, fechapublicacion__isnull=False).annotate(anio=ExtractYear('fechapublicacion')).values_list('anio', flat=True).order_by('-anio').distinct()

                for anio in anios:
                    print("Procesando año ", anio)
                    ponencias_publicadas = PonenciasInvestigacion.objects.filter(status=True, fechapublicacion__year=anio, pk__in=lista_ids).order_by('id')
                    ponencias_no_publicadas = PonenciasInvestigacion.objects.filter(status=True, fechapublicacion__isnull=True, fechainicio__year=anio, pk__in=lista_ids).order_by('id')

                    total = ponencias_publicadas.count() + ponencias_no_publicadas.count()
                    rp = 0

                    if ponencias_publicadas or ponencias_no_publicadas:
                        print("Publicados: ", ponencias_publicadas.count())
                        print("No Publicados: ", ponencias_no_publicadas.count())
                        for ponencia in ponencias_publicadas:
                            rp += 1
                            print("Procesando", rp, " de ", total)

                            titulo = ponencia.nombre

                            palabras = titulo.split(" ")
                            titulo = "_".join(palabras[0:5])
                            titulo = remover_caracteres(titulo, caracteres_a_quitar)

                            titulo = elimina_tildes(remover_caracteres_tildes_unicode(remover_caracteres_especiales_unicode(titulo)))
                            carpetaponencia = "ANIO_" + str(anio) + "/PUBLICADOS/PON_" + str(ponencia.id) + "_" + titulo

                            ponencia_id = ponencia.id
                            nombre = "PONENCIA_" + str(ponencia_id).zfill(4)

                            # Agrego las evidencias a la carpeta de la ponencia
                            for evidencia in ponencia.detalleevidencias_set.filter(status=True):
                                ext = evidencia.archivo.__str__()[evidencia.archivo.__str__().rfind("."):]

                                if evidencia.descripcion:
                                    nombreevidencia = evidencia.descripcion.upper()
                                    nombreevidencia = remover_caracteres(nombreevidencia, caracteres_a_quitar)
                                    nombreevidencia = elimina_tildes(remover_caracteres_tildes_unicode(remover_caracteres_especiales_unicode(nombreevidencia))).replace(" ", "_")
                                else:
                                    nombreevidencia = "EVIDENCIA" + random.randint(1, 10000).__str__()


                                if evidencia.evidencia.id == 4:
                                    if os.path.exists(SITE_STORAGE + evidencia.archivo.url):
                                        fantasy_zip.write(SITE_STORAGE + evidencia.archivo.url, carpetaponencia + '/' + nombreevidencia + ext.lower())

                                elif evidencia.evidencia.id == 5:
                                    if os.path.exists(SITE_STORAGE + evidencia.archivo.url):
                                        fantasy_zip.write(SITE_STORAGE + evidencia.archivo.url, carpetaponencia + '/' + nombreevidencia + ext.lower())

                                elif evidencia.evidencia.id == 6:
                                    if os.path.exists(SITE_STORAGE + evidencia.archivo.url):
                                        fantasy_zip.write(SITE_STORAGE + evidencia.archivo.url, carpetaponencia + '/' + nombreevidencia + ext.lower())

                                elif evidencia.evidencia.id == 15:
                                    if os.path.exists(SITE_STORAGE + evidencia.archivo.url):
                                        fantasy_zip.write(SITE_STORAGE + evidencia.archivo.url, carpetaponencia + '/' + nombreevidencia + ext.lower())

                            # Generar la ficha catalográfica de la ponencia
                            data['ponencia'] = ponencia
                            data['participantes'] = ponencia.participantes()

                            nombrearchivoficha = 'fichacatalograficapon_' + str(ponencia_id) + '_' + str(random.randint(1, 10000)) + '.pdf'
                            valida = convert_html_to_pdf(
                                'inv_ponencia/fichacatalografica_pdf.html',
                                {'pagesize': 'A4', 'data': data},
                                nombrearchivoficha,
                                directorio
                            )

                            archivoficha = directorio + "/" + nombrearchivoficha
                            # Agrego el archivo de la ficha a la carpeta del artículo
                            fantasy_zip.write(archivoficha, carpetaponencia + "/FICHA_CATALOGRAFICA.pdf")
                            # Borro el archivo de la ficha creado
                            os.remove(archivoficha)

                        for ponencia in ponencias_no_publicadas:
                            rp += 1
                            print("Procesando", rp, " de ", total)

                            titulo = ponencia.nombre

                            palabras = titulo.split(" ")
                            titulo = "_".join(palabras[0:5])
                            titulo = remover_caracteres(titulo, caracteres_a_quitar)

                            titulo = elimina_tildes(remover_caracteres_tildes_unicode(remover_caracteres_especiales_unicode(titulo)))
                            carpetaponencia = "ANIO_" + str(anio) + "/NOPUBLICADAS/PON_" + str(ponencia.id) + "_" + titulo

                            ponencia_id = ponencia.id
                            nombre = "PONENCIA_" + str(ponencia_id).zfill(4)

                            # Agrego las evidencias a la carpeta de la ponencia
                            for evidencia in ponencia.detalleevidencias_set.filter(status=True):
                                ext = evidencia.archivo.__str__()[evidencia.archivo.__str__().rfind("."):]

                                if evidencia.descripcion:
                                    nombreevidencia = evidencia.descripcion.upper()
                                    nombreevidencia = remover_caracteres(nombreevidencia, caracteres_a_quitar)
                                    nombreevidencia = elimina_tildes(remover_caracteres_tildes_unicode(remover_caracteres_especiales_unicode(nombreevidencia))).replace(" ", "_")
                                else:
                                    nombreevidencia = "EVIDENCIA" + random.randint(1, 10000).__str__()

                                if evidencia.evidencia.id == 4:
                                    if os.path.exists(SITE_STORAGE + evidencia.archivo.url):
                                        fantasy_zip.write(SITE_STORAGE + evidencia.archivo.url, carpetaponencia + '/' + nombreevidencia + ext.lower())

                                elif evidencia.evidencia.id == 5:
                                    if os.path.exists(SITE_STORAGE + evidencia.archivo.url):
                                        fantasy_zip.write(SITE_STORAGE + evidencia.archivo.url, carpetaponencia + '/' + nombreevidencia + ext.lower())

                                elif evidencia.evidencia.id == 6:
                                    if os.path.exists(SITE_STORAGE + evidencia.archivo.url):
                                        fantasy_zip.write(SITE_STORAGE + evidencia.archivo.url, carpetaponencia + '/' + nombreevidencia + ext.lower())

                                elif evidencia.evidencia.id == 15:
                                    if os.path.exists(SITE_STORAGE + evidencia.archivo.url):
                                        fantasy_zip.write(SITE_STORAGE + evidencia.archivo.url, carpetaponencia + '/' + nombreevidencia + ext.lower())

                            # Generar la ficha catalográfica de la ponencia
                            data['ponencia'] = ponencia
                            data['participantes'] = ponencia.participantes()

                            nombrearchivoficha = 'fichacatalograficapon_' + str(ponencia_id) + '_' + str(random.randint(1, 10000)) + '.pdf'
                            valida = convert_html_to_pdf(
                                'inv_ponencia/fichacatalografica_pdf.html',
                                {'pagesize': 'A4', 'data': data},
                                nombrearchivoficha,
                                directorio
                            )

                            archivoficha = directorio + "/" + nombrearchivoficha
                            # Agrego el archivo de la ficha a la carpeta del artículo
                            fantasy_zip.write(archivoficha, carpetaponencia + "/FICHA_CATALOGRAFICA.pdf")
                            # Borro el archivo de la ficha creado
                            os.remove(archivoficha)

                fantasy_zip.close()
                ruta = "media/zipav/" + nombre_archivo
                return JsonResponse({'result': 'ok', 'archivo': ruta})
            except Exception as ex:
                msg = ex.__str__()
                return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al descargar evidencias. Detalle: [%s]" % msg, "showSwal": "True", "swalType": "error"})


        return JsonResponse({"result": "bad", "mensaje": u"Solicitud Incorrecta."})

    else:
        if 'action' in request.GET:
            action = request.GET['action']

            if action == 'aprobar':
                try:
                    data['title'] = u'Aprobar Solicitud Ponencias'
                    data['solicitud'] = SolicitudPublicacion.objects.get(pk=request.GET['id'])
                    return render(request, "inv_ponencia/aprobar.html", data)
                except Exception as ex:
                    pass

            elif action == 'aprobarponencia':
                try:
                    data['title'] = u'Aprobar Ponencias'
                    ponencia = PonenciasInvestigacion.objects.get(pk=request.GET['idponencia'])
                    data['ponencia'] = ponencia
                    # data['solicitud'] = SolicitudPublicacion.objects.get(pk=ponencia.solicitudpublicacion.id)
                    return render(request, "inv_ponencia/aprobar.html", data)
                except Exception as ex:
                    pass

            elif action == 'add':
                try:
                    data['title'] = u'Adicionar Ponencias (Sin Solicitud)'
                    form = PonenciaInvestigacionForm()
                    data['form'] = form
                    return render(request, "inv_ponencia/add.html", data)
                except Exception as ex:
                    pass

            elif action == 'mostrarevidenciaregistro':
                try:
                    data['title'] = u'Evidencias de Registro de Ponencia'
                    data['solicitud'] = SolicitudPublicacion.objects.get(pk=int(encrypt(request.GET['idrp'])))
                    template = get_template("inv_ponencia/mostrarevidenciaregistroponencia.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content, 'title': data['title']})
                except Exception as ex:
                    pass

            elif action == 'ingresarsolicitudponencia':
                try:
                    data['title'] = u'Adicionar Ponencias'
                    solicitud = SolicitudPublicacion.objects.get(pk=int(encrypt(request.GET['id'])))

                    form = PonenciaInvestigacionRegistroForm(initial={
                                                        'nombre': solicitud.nombre,
                                                        'resumen': solicitud.motivo,
                                                        'evento': solicitud.evento,
                                                        'pais': solicitud.pais,
                                                        'ciudad': solicitud.ciudad,
                                                        'fechainicio': solicitud.fecharecepcion,
                                                        'fechafin': solicitud.fechaaprobacion,
                                                        'enlace': solicitud.enlace,
                                                        'estadopublicacion': solicitud.estadopublicacion,
                                                        'fechapublicacion': solicitud.fechapublicacion,
                                                        'areaconocimiento': solicitud.areaconocimiento,
                                                        'subareaconocimiento': solicitud.subareaconocimiento,
                                                        'subareaespecificaconocimiento': solicitud.subareaespecificaconocimiento,
                                                        'lineainvestigacion': solicitud.lineainvestigacion,
                                                        'sublineainvestigacion': solicitud.sublineainvestigacion,
                                                        'provieneproyecto': solicitud.provieneproyecto,
                                                        'tipoproyecto': solicitud.tipoproyecto,
                                                        'proyectointerno': solicitud.proyectointerno,
                                                        'proyectoexterno': solicitud.proyectoexterno,
                                                        'pertenecegrupoinv': solicitud.pertenecegrupoinv,
                                                        'grupoinvestigacion': solicitud.grupoinvestigacion,
                                                        'comitecientifico': solicitud.comitecientifico
                    })

                    form.editar(solicitud)
                    data['form'] = form
                    data['idsolicitud'] = solicitud.id
                    data['integrantescomite'] = integrantes = solicitud.integrantes_comite_cientifico_ponencia()
                    data['totalintegrantes'] = len(integrantes)
                    data['evidencias'] = solicitud.evidencias()
                    data['participantes'] = solicitud.participantes()
                    data['estadossolicitud'] = obtener_estados_solicitud(8, [2, 3, 4])

                    return render(request, "inv_ponencia/ingresarsolicitudponencia.html", data)
                except Exception as ex:
                    pass

            elif action == 'editarponencia':
                try:
                    data['title'] = u'Editar Ponencia'
                    data['ponencias'] = ponencia = PonenciasInvestigacion.objects.get(pk=int(encrypt(request.GET['id'])))
                    form = PonenciaInvestigacionRegistroForm(initial={'nombre': ponencia.nombre,
                                                              'evento': ponencia.evento,
                                                              'resumen': ponencia.resumen,
                                                              # 'estado': ponencias.estado,
                                                              'pais': ponencia.pais,
                                                              'ciudad': ponencia.ciudad,
                                                              'fechainicio': ponencia.fechainicio,
                                                              'fechafin': ponencia.fechafin,
                                                              'enlace': ponencia.enlace,
                                                              'estadopublicacion': ponencia.estado,
                                                              'fechapublicacion': ponencia.fechapublicacion,
                                                              'areaconocimiento': ponencia.areaconocimiento,
                                                              'subareaconocimiento': ponencia.subareaconocimiento,
                                                              'subareaespecificaconocimiento': ponencia.subareaespecificaconocimiento,
                                                              'lineainvestigacion': ponencia.lineainvestigacion,
                                                              'sublineainvestigacion': ponencia.sublineainvestigacion,
                                                              'provieneproyecto': ponencia.provieneproyecto,
                                                              'tipoproyecto': ponencia.tipoproyecto,
                                                              'proyectointerno': ponencia.proyectointerno,
                                                              'proyectoexterno': ponencia.proyectoexterno,
                                                              'pertenecegrupoinv': ponencia.pertenecegrupoinv,
                                                              'grupoinvestigacion': ponencia.grupoinvestigacion,
                                                              'comitecientifico': ponencia.comitecientifico,
                                                              'numerocongreso' : ponencia.numerocongreso,
                                                              'organizadorevento' : ponencia.organizadorevento,
                                                              'comiteorganizador' : ponencia.comiteorganizador,
                                                              'accesoabierto' : ponencia.accesoabierto
                                                              })
                    form.editar(ponencia)
                    data['form'] = form
                    integrantescomite = None
                    if ponencia.integrantecomite():
                        integrantescomite = "|".join([i.integrante+","+i.institucion+","+i.email for i in ponencia.integrantecomite()])
                    data['integrantescomite'] = integrantescomite

                    # data['mostrarguardar'] = False if ponencia.aprobado else True
                    data['mostrarguardar'] = True

                    return render(request, "inv_ponencia/editarponencia.html", data)
                except Exception as ex:
                    pass

            elif action == 'deleteparticipanteponencia':
                try:
                    data['title'] = u'Eliminar Participante'
                    tipo = request.GET['tipo']
                    data['participante'] = participante = ParticipantePonencias.objects.get(pk=request.GET['id'])
                    if tipo == '1':
                        data['nombres'] = participante.profesor.persona.nombre_completo()
                    if tipo == '3':
                        data['nombres'] = participante.administrativo.persona.nombre_completo()
                    if tipo == '4':
                        data['nombres'] = participante.inscripcion.persona.nombre_completo()

                    return render(request, "inv_ponencia/deleteparticipanteponencia.html", data)
                except Exception as ex:
                    pass

            elif action == 'deleteponencia':
                try:
                    data['title'] = u'Eliminar Ponencia'
                    data['ponencias'] = PonenciasInvestigacion.objects.get(pk=request.GET['idponencia'])
                    return render(request, "inv_ponencia/deleteponencia.html", data)
                except Exception as ex:
                    pass

            elif action == 'participantesponencia':
                try:
                    data['title'] = u'Participantes de Ponencias'
                    data['ponencias'] = ponencias = PonenciasInvestigacion.objects.get(pk=int(encrypt(request.GET['id'])))
                    data['tipoparticipante'] = TIPO_PARTICIPANTE
                    data['tipoparinstitucion'] = TIPO_PARTICIPANTE_INSTITUCION
                    data['participantes'] = ParticipantePonencias.objects.filter(status=True, ponencias=ponencias)
                    return render(request, "inv_ponencia/participantesponencia.html", data)
                except Exception as ex:
                    pass

            elif action == 'addparticipantesdocentes':
                try:
                    data['title'] = u'Participante Docente'
                    data['form'] = ParticipanteProfesorPonenciaForm
                    data['id'] = request.GET['idponencia']
                    return render(request, "inv_ponencia/addparticipantedocente.html", data)
                except Exception as ex:
                    pass

            elif action == 'addparticipantesadministrativos':
                try:
                    data['title'] = u'Participante Administrativo'
                    data['form'] = ParticipanteAdministrativoPonenciaForm
                    data['id'] = request.GET['idponencia']
                    return render(request, "inv_ponencia/addparticipanteadministrativo.html", data)
                except Exception as ex:
                    pass

            elif action == 'addparticipantesinscripcion':
                try:
                    data['title'] = u'Participante Estudiante'
                    data['form'] = ParticipanteInscripcionPonenciaForm
                    data['id'] = request.GET['idponencia']
                    return render(request, "inv_ponencia/addparticipanteinscripcion.html", data)
                except Exception as ex:
                    pass

            elif action == 'evidenciasponencia':
                try:
                    data['title'] = u'Evidencia Ponencias'
                    data['ponencias'] = ponencia = PonenciasInvestigacion.objects.get(pk=int(encrypt(request.GET['id'])))
                    data['evidencias'] = Evidencia.objects.filter(status=True, matrizevidencia_id=4) if ponencia.comitecientifico else Evidencia.objects.filter(status=True, matrizevidencia_id=4).exclude(pk=15)
                    data['formevidencias'] = EvidenciaForm()
                    return render(request, "inv_ponencia/evidenciasponencias.html", data)
                except Exception as ex:
                    pass

            elif action == 'addevidenciasponencia':
                try:
                    evidencia = Evidencia.objects.get(pk=int(encrypt(request.GET['idevidencia'])))
                    detalleevidencia = DetalleEvidencias.objects.filter(evidencia_id=int(encrypt(request.GET['idevidencia'])), ponencia_id=int(encrypt(request.GET['id'])))
                    data['title'] = u'Editar Evidencia'
                    data['evidencia'] = evidencia.__str__().upper()

                    if detalleevidencia:
                        detalleevidencia = detalleevidencia[0]
                        form = EvidenciaForm(initial={'descripcion': detalleevidencia.descripcion})
                    else:
                        if evidencia.id == 4:
                            descripcion = "CARTA DE ACEPTACIÓN"
                        elif evidencia.id == 5:
                            descripcion = "CERTIFICADO DE PARTICIPACIÓN"
                        elif evidencia.id == 6:
                            descripcion = "PONENCIA PUBLICADA"
                        elif evidencia.id == 15:
                            descripcion = "COMITÉ CIENTÍFICO"
                        else:
                            descripcion = "PROGRAMA DEL EVENTO"

                        form = EvidenciaForm(initial={'descripcion': descripcion})

                    data['form'] = form
                    data['id'] = request.GET['id']
                    data['idevidencia'] = request.GET['idevidencia']

                    template = get_template("inv_ponencia/add_evidenciasponencias.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content})
                except Exception as ex:
                    pass

            elif action == 'rechazar':
                try:
                    data['title'] = u'Rechazar Solicitud'
                    data['ponencia'] = SolicitudPublicacion.objects.get(pk=request.GET['id'])
                    return render(request, "inv_ponencia/rechazar.html", data)
                except Exception as ex:
                    pass

            elif action == 'planificarponencias':
                try:
                    data['reporte_planificar_ponencia'] = obtener_reporte('reporte_planificar_ponencia')
                    data['title'] = u'Listado de Solicitudes de Financiamiento a ponencias'
                    ponencias = PlanificarPonencias.objects.filter(status=True).order_by('-fecha_creacion')
                    paging = Paginator(ponencias, 30)
                    p = 1
                    try:
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        page = paging.page(p)
                    except Exception as ex:
                        page = paging.page(1)
                    data['paging'] = paging
                    data['page'] = page
                    data['ponencias'] = page.object_list
                    return render(request, "inv_ponencia/planificarponencias.html", data)
                except Exception as ex:
                    pass

            elif action == 'reporte_ponencias_excel':
                try:
                    __author__ = 'Unemi'
                    fuentenormal = easyxtitulo = easyxf(
                        'font: name Times New Roman, color-index black, bold on , height 350; alignment: horiz centre')
                    titulo = easyxf(
                        'font: name Times New Roman, color-index black, bold on , height 350; alignment: horiz centre')
                    titulo2 = easyxf(
                        'font: name Times New Roman, color-index black, bold on , height 250; alignment: horiz centre')
                    fuentecabecera = easyxf(
                        'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
                    fuentenormal = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
                    fuentenormalwrap = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
                    fuentenormalwrap.alignment.wrap = True
                    fuentenormalcent = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz centre')
                    fuentemoneda = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right',
                        num_format_str=' "$" #,##0.00')
                    fuentefecha = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz center',
                        num_format_str='yyyy-mm-dd')
                    fuentenumerodecimal = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right',
                        num_format_str='#,##0.00')
                    fuentenumeroentero = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right')

                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('Listado')
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=lista_ponencias_' + random.randint(1, 10000).__str__() + '.xls'

                    ws.write_merge(0, 0, 0, 25, 'UNIVERSIDAD ESTATAL DE MILAGRO', titulo)
                    ws.write_merge(1, 1, 0, 25, 'VICERRECTORADO DE INVESTIGACIÓN Y POSGRADO', titulo2)
                    ws.write_merge(2, 2, 0, 25, 'COORDINACIÓN DE INVESTIGACIÓN', titulo2)
                    ws.write_merge(3, 3, 0, 25, 'LISTADO DE PONENCIAS', titulo2)

                    row_num = 5
                    columns = [
                        (u"TIPO", 2500),
                        (u"CÓDIGO", 4500),
                        (u"NOMBRE DE PONENCIA", 10000),
                        (u"NOMBRE DEL EVENTO", 10000),
                        (u"PAÍS", 5000),
                        (u"CIUDAD", 5000),
                        (u"FECHA PUBLICACIÓN", 5000),
                        (u"ÁREA DE CONOCIMIENTO", 5000),
                        (u"SUB-ÁREA DE CONOCIMIENTO", 5000),
                        (u"SUB-ÁREA ESPECÍFICA", 5000),
                        (u"LÍNEA INVESTIGACIÓN", 5000),
                        (u"SUB-LÍNEA INVESTIGACIÓN", 5000),
                        (u"PROVIENE DE PROYECTO", 3000),
                        (u"TIPO PROYECTO", 3000),
                        (u"TÍTULO DEL PROYECTO", 10000),
                        (u"PERTENECE GRUPO INVESTIGACIÓN", 3000),
                        (u"GRUPO DE INVESTIGACIÓN", 10000),
                        (u"CÉDULA", 3000),
                        (u"TIPO PARTICIPANTE", 3000),
                        (u"PARTICIPANTE", 8000),
                        (u"TIPO PARTICIPACIÓN", 3000),
                        (u"TIPO UNEMI", 3000),
                        (u"EDICIÓN DEL EVENTO", 8000),
                        (u"COMITÉ CIENTÍFICO", 8000),
                        (u"ORGANIZADOR DEL EVENTO", 8000),
                        (u"COMITÉ ORGANIZADOR", 8000)
                    ]

                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], fuentecabecera)
                        ws.col(col_num).width = columns[col_num][1]

                    listaponencias = ParticipantePonencias.objects.select_related().filter(ponencias__status=True, status=True).order_by('ponencias__nombre')

                    for ponencia in listaponencias:
                        row_num += 1

                        if ponencia.profesor:
                            nombres = ponencia.profesor.persona.nombre_completo_inverso()
                            cedula = ponencia.profesor.persona.cedula
                        elif ponencia.administrativo:
                            nombres = ponencia.administrativo.persona.nombre_completo_inverso()
                            cedula = ponencia.administrativo.persona.cedula
                        elif ponencia.inscripcion:
                            nombres = ponencia.inscripcion.persona.nombre_completo_inverso()
                            cedula = ponencia.inscripcion.persona.cedula

                        ws.write(row_num, 0, "PONENCIA", fuentenormal)
                        ws.write(row_num, 1, str(ponencia.ponencias.id) + '-PON', fuentenormal)
                        ws.write(row_num, 2, ponencia.ponencias.nombre.strip(), fuentenormal)
                        ws.write(row_num, 3, ponencia.ponencias.evento.strip(), fuentenormal)
                        ws.write(row_num, 4, ponencia.ponencias.pais.nombre, fuentenormal)
                        ws.write(row_num, 5, ponencia.ponencias.ciudad, fuentenormal)
                        ws.write(row_num, 6, ponencia.ponencias.fechapublicacion, fuentefecha)
                        ws.write(row_num, 7, ponencia.ponencias.areaconocimiento.nombre, fuentenormal)
                        ws.write(row_num, 8, ponencia.ponencias.subareaconocimiento.nombre, fuentenormal)
                        ws.write(row_num, 9, ponencia.ponencias.subareaespecificaconocimiento.nombre, fuentenormal)
                        ws.write(row_num, 10, ponencia.ponencias.lineainvestigacion.nombre if ponencia.ponencias.lineainvestigacion else '', fuentenormal)
                        ws.write(row_num, 11, ponencia.ponencias.sublineainvestigacion.nombre if ponencia.ponencias.sublineainvestigacion else '', fuentenormal)
                        ws.write(row_num, 12, 'SI' if ponencia.ponencias.tipoproyecto else 'NO', fuentenormal)
                        ws.write(row_num, 13, ponencia.ponencias.get_tipoproyecto_display() if ponencia.ponencias.tipoproyecto else '', fuentenormal)
                        if ponencia.ponencias.tipoproyecto:
                            ws.write(row_num, 14, ponencia.ponencias.proyectointerno.nombre if ponencia.ponencias.proyectointerno else ponencia.ponencias.proyectoexterno.nombre, fuentenormal)
                        else:
                            ws.write(row_num, 14, '', fuentenormal)

                        ws.write(row_num, 15, 'SI' if ponencia.ponencias.pertenecegrupoinv else 'NO', fuentenormal)
                        ws.write(row_num, 16, ponencia.ponencias.grupoinvestigacion.nombre if ponencia.ponencias.pertenecegrupoinv else '', fuentenormal)
                        ws.write(row_num, 17, cedula, fuentenormal)
                        ws.write(row_num, 18, ponencia.get_tipoparticipante_display(), fuentenormal)
                        ws.write(row_num, 19, nombres, fuentenormal)
                        ws.write(row_num, 20, ponencia.get_tipoparticipanteins_display(), fuentenormal)
                        ws.write(row_num, 21, ponencia.tipo_unemi(), fuentenormal)
                        ws.write(row_num, 22, ponencia.ponencias.numerocongreso, fuentenormal)
                        ws.write(row_num, 23, ponencia.ponencias.nombrecomite, fuentenormal)
                        ws.write(row_num, 24, ponencia.ponencias.organizadorevento, fuentenormal)
                        ws.write(row_num, 25, ponencia.ponencias.comiteorganizador, fuentenormal)

                    wb.save(response)
                    return response
                except Exception as ex:
                    pass

            elif action == 'reporte_ponencias_excel_participante':
                try:
                    __author__ = 'Unemi'
                    fuentenormal = easyxtitulo = easyxf(
                        'font: name Times New Roman, color-index black, bold on , height 350; alignment: horiz centre')
                    titulo = easyxf(
                        'font: name Times New Roman, color-index black, bold on , height 350; alignment: horiz centre')
                    titulo2 = easyxf(
                        'font: name Times New Roman, color-index black, bold on , height 250; alignment: horiz centre')
                    fuentecabecera = easyxf(
                        'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
                    fuentenormal = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
                    fuentenormalwrap = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
                    fuentenormalwrap.alignment.wrap = True
                    fuentenormalcent = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz centre')
                    fuentemoneda = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right',
                        num_format_str=' "$" #,##0.00')
                    fuentefecha = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz center',
                        num_format_str='yyyy-mm-dd')
                    fuentenumerodecimal = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right',
                        num_format_str='#,##0.00')
                    fuentenumeroentero = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right')

                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('Listado')
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=lista_ponencias_participante_' + random.randint(1, 10000).__str__() + '.xls'

                    ws.write_merge(0, 0, 0, 25, 'UNIVERSIDAD ESTATAL DE MILAGRO', titulo)
                    ws.write_merge(1, 1, 0, 25, 'VICERRECTORADO DE INVESTIGACIÓN Y POSGRADO', titulo2)
                    ws.write_merge(2, 2, 0, 25, 'COORDINACIÓN DE INVESTIGACIÓN', titulo2)
                    ws.write_merge(3, 3, 0, 25, 'LISTADO DE PONENCIAS POR PARTICIPANTE', titulo2)

                    row_num = 5

                    columns = [
                        (u"TIPO", 2500),
                        (u"CÓDIGO", 4500),
                        (u"NOMBRE DE PONENCIA", 10000),
                        (u"NOMBRE DEL EVENTO", 10000),
                        (u"PAÍS", 5000),
                        (u"CIUDAD", 5000),
                        (u"FECHA PUBLICACIÓN", 5000),
                        (u"ÁREA DE CONOCIMIENTO", 5000),
                        (u"SUB-ÁREA DE CONOCIMIENTO", 5000),
                        (u"SUB-ÁREA ESPECÍFICA", 5000),
                        (u"LÍNEA INVESTIGACIÓN", 5000),
                        (u"SUB-LÍNEA INVESTIGACIÓN", 5000),
                        (u"PROVIENE DE PROYECTO", 3000),
                        (u"TIPO PROYECTO", 3000),
                        (u"TÍTULO DEL PROYECTO", 10000),
                        (u"PERTENECE GRUPO INVESTIGACIÓN", 3000),
                        (u"GRUPO DE INVESTIGACIÓN", 10000),
                        (u"CÉDULA", 3000),
                        (u"TIPO PARTICIPANTE", 3000),
                        (u"PARTICIPANTE", 8000),
                        (u"TIPO PARTICIPACIÓN", 3000),
                        (u"TIPO UNEMI", 3000),
                        (u"EDICIÓN DEL EVENTO", 8000),
                        (u"COMITÉ CIENTÍFICO", 8000),
                        (u"ORGANIZADOR DEL EVENTO", 8000),
                        (u"COMITÉ ORGANIZADOR", 8000)
                    ]

                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], fuentecabecera)
                        ws.col(col_num).width = columns[col_num][1]

                    lista_ids_personas = request.GET['idsp'].split(",")

                    listaponencias = ParticipantePonencias.objects.select_related().filter(Q(profesor__persona__id__in=lista_ids_personas)|
                                                                                           Q(administrativo__persona__id__in=lista_ids_personas)|
                                                                                           Q(inscripcion__persona__id__in=lista_ids_personas),
                                                                                           ponencias__status=True, status=True).order_by('ponencias__nombre')

                    for ponencia in listaponencias:
                        row_num += 1

                        if ponencia.profesor:
                            nombres = ponencia.profesor.persona.nombre_completo_inverso()
                            cedula = ponencia.profesor.persona.cedula
                        elif ponencia.administrativo:
                            nombres = ponencia.administrativo.persona.nombre_completo_inverso()
                            cedula = ponencia.administrativo.persona.cedula
                        elif ponencia.inscripcion:
                            nombres = ponencia.inscripcion.persona.nombre_completo_inverso()
                            cedula = ponencia.inscripcion.persona.cedula

                        ws.write(row_num, 0, "PONENCIA", fuentenormal)
                        ws.write(row_num, 1, str(ponencia.ponencias.id) + '-PON', fuentenormal)
                        ws.write(row_num, 2, ponencia.ponencias.nombre.strip(), fuentenormal)
                        ws.write(row_num, 3, ponencia.ponencias.evento.strip(), fuentenormal)
                        ws.write(row_num, 4, ponencia.ponencias.pais.nombre, fuentenormal)
                        ws.write(row_num, 5, ponencia.ponencias.ciudad, fuentenormal)
                        ws.write(row_num, 6, ponencia.ponencias.fechapublicacion, fuentefecha)
                        ws.write(row_num, 7, ponencia.ponencias.areaconocimiento.nombre, fuentenormal)
                        ws.write(row_num, 8, ponencia.ponencias.subareaconocimiento.nombre, fuentenormal)
                        ws.write(row_num, 9, ponencia.ponencias.subareaespecificaconocimiento.nombre, fuentenormal)
                        ws.write(row_num, 10, ponencia.ponencias.lineainvestigacion.nombre if ponencia.ponencias.lineainvestigacion else '', fuentenormal)
                        ws.write(row_num, 11, ponencia.ponencias.sublineainvestigacion.nombre if ponencia.ponencias.sublineainvestigacion else '', fuentenormal)
                        ws.write(row_num, 12, 'SI' if ponencia.ponencias.tipoproyecto else 'NO', fuentenormal)
                        ws.write(row_num, 13, ponencia.ponencias.get_tipoproyecto_display() if ponencia.ponencias.tipoproyecto else '', fuentenormal)
                        if ponencia.ponencias.tipoproyecto:
                            ws.write(row_num, 14, ponencia.ponencias.proyectointerno.nombre if ponencia.ponencias.proyectointerno else ponencia.ponencias.proyectoexterno.nombre, fuentenormal)
                        else:
                            ws.write(row_num, 14, '', fuentenormal)

                        ws.write(row_num, 15, 'SI' if ponencia.ponencias.pertenecegrupoinv else 'NO', fuentenormal)
                        ws.write(row_num, 16, ponencia.ponencias.grupoinvestigacion.nombre if ponencia.ponencias.pertenecegrupoinv else '', fuentenormal)
                        ws.write(row_num, 17, cedula, fuentenormal)
                        ws.write(row_num, 18, ponencia.get_tipoparticipante_display(), fuentenormal)
                        ws.write(row_num, 19, nombres, fuentenormal)
                        ws.write(row_num, 20, ponencia.get_tipoparticipanteins_display(), fuentenormal)
                        ws.write(row_num, 21, ponencia.tipo_unemi(), fuentenormal)
                        ws.write(row_num, 22, ponencia.ponencias.numerocongreso, fuentenormal)
                        ws.write(row_num, 23, ponencia.ponencias.nombrecomite, fuentenormal)
                        ws.write(row_num, 24, ponencia.ponencias.organizadorevento, fuentenormal)
                        ws.write(row_num, 25, ponencia.ponencias.comiteorganizador, fuentenormal)

                    wb.save(response)
                    return response
                except Exception as ex:
                    pass

            # elif action == 'excelponencias':
            #     try:
            #         __author__ = 'Unemi'
            #         style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
            #         style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
            #         style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
            #         title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
            #         style1 = easyxf(num_format_str='D-MMM-YY')
            #         font_style = XFStyle()
            #         font_style.font.bold = True
            #         font_style2 = XFStyle()
            #         font_style2.font.bold = False
            #         wb = Workbook(encoding='utf-8')
            #         ws = wb.add_sheet('exp_xls_post_part')
            #         ws.write_merge(0, 0, 0, 12, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
            #         response = HttpResponse(content_type="application/ms-excel")
            #         response[
            #             'Content-Disposition'] = 'attachment; filename=Listas_Ponencias' + random.randint(
            #             1, 10000).__str__() + '.xls'
            #
            #         columns = [
            #             (u"TIPO", 2500),
            #             (u"CODIGO", 4500),
            #             (u"NOMBRE DE PONENCIA", 10000),
            #             (u"NOMBRE DEL EVENTO", 10000),
            #             (u"PAIS", 5000),
            #             (u"CIUDAD", 5000),
            #             (u"FECHA PUBLICACIÓN", 5000),
            #             (u"AREA DE CONOCIMIENTO", 5000),
            #             (u"SUBAREA DE CONOCIMIENTO", 5000),
            #             (u"SUBAREA ESPECÍFICA", 5000),
            #             (u"CEDULA", 3000),
            #             (u"TIPO PARTICIPANTE", 3000),
            #             (u"PARTICIPANTE", 8000),
            #
            #
            #             (u"TIPO PARTICIPACIÓN", 3000),
            #             (u"TIPO UNEMI", 3000),
            #
            #
            #             (u"EDICIÓN DEL EVENTO", 8000),
            #             (u"COMITÉ CIENTÍFICO", 8000),
            #             (u"ORGANIZADOR DEL EVENTO", 8000),
            #             (u"COMITÉ ORGANIZADOR", 8000),
            #         ]
            #         row_num = 3
            #         for col_num in  range(len(columns)):
            #             ws.write(row_num, col_num, columns[col_num][0], font_style)
            #             ws.col(col_num).width = columns[col_num][1]
            #         date_format = xlwt.XFStyle()
            #         date_format.num_format_str = 'yyyy/mm/dd'
            #         listaponencias = ParticipantePonencias.objects.select_related().filter(ponencias__status=True, status=True).order_by('ponencias__nombre')
            #         row_num = 4
            #         for ponencia in listaponencias:
            #             i = 0
            #             campo10 = None
            #             nombre = ponencia.ponencias.nombre
            #             evento = ponencia.ponencias.evento
            #             pais = ponencia.ponencias.pais.nombre
            #             ciudad = ponencia.ponencias.ciudad
            #             fechapublica = ponencia.ponencias.fechapublicacion
            #             area = ponencia.ponencias.areaconocimiento.nombre
            #             subarea = ponencia.ponencias.subareaconocimiento.nombre
            #             subareaespecifica = ponencia.ponencias.subareaespecificaconocimiento.nombre
            #             tipoparticipacion = ponencia.get_tipoparticipanteins_display()
            #             tipounemi = ponencia.tipo_unemi()
            #             tipoparticipante = ponencia.get_tipoparticipante_display()
            #
            #             if ponencia.profesor:
            #                 nombres = ponencia.profesor.persona.apellido1 + ' ' + ponencia.profesor.persona.apellido2 + ' ' + ponencia.profesor.persona.nombres
            #                 cedula = ponencia.profesor.persona.cedula
            #             if ponencia.administrativo:
            #                 nombres = ponencia.administrativo.persona.apellido1 + ' ' + ponencia.administrativo.persona.apellido2 + ' ' + ponencia.administrativo.persona.nombres
            #                 cedula = ponencia.administrativo.persona.cedula
            #             if ponencia.inscripcion:
            #                 nombres = ponencia.inscripcion.persona.apellido1 + ' ' + ponencia.inscripcion.persona.apellido2 + ' ' + ponencia.inscripcion.persona.nombres
            #                 cedula = ponencia.inscripcion.persona.cedula
            #
            #             numerocongreso = ponencia.ponencias.numerocongreso
            #             nombrecomite = ponencia.ponencias.nombrecomite
            #             organizadorevento = ponencia.ponencias.organizadorevento
            #             comiteorganizador = ponencia.ponencias.comiteorganizador
            #             ws.write(row_num, 0, 'PONENCIA', font_style2)
            #             ws.write(row_num, 1, str(ponencia.ponencias.id) + '-PON', font_style2)
            #             ws.write(row_num, 2, nombre, font_style2)
            #             ws.write(row_num, 3, evento, font_style2)
            #             ws.write(row_num, 4, pais, font_style2)
            #             ws.write(row_num, 5, ciudad, font_style2)
            #             ws.write(row_num, 6, fechapublica, date_format)
            #             ws.write(row_num, 7, area, font_style2)
            #             ws.write(row_num, 8, subarea, font_style2)
            #             ws.write(row_num, 9, subareaespecifica, font_style2)
            #             ws.write(row_num, 10, cedula, font_style2)
            #             ws.write(row_num, 11, tipoparticipante, font_style2)
            #             ws.write(row_num, 12, nombres, font_style2)
            #             ws.write(row_num, 13, tipoparticipacion, font_style2)
            #             ws.write(row_num, 14, tipounemi, font_style2)
            #             ws.write(row_num, 15, numerocongreso, font_style2)
            #             ws.write(row_num, 16, nombrecomite, font_style2)
            #             ws.write(row_num, 17, organizadorevento, font_style2)
            #             ws.write(row_num, 18, comiteorganizador, font_style2)
            #             row_num += 1
            #         wb.save(response)
            #         return response
            #     except Exception as ex:
            #         pass

            elif action == 'editponencia':
                try:
                    data['title'] = u'Editar planificación de ponencias'
                    planificarponencias = PlanificarPonencias.objects.get(pk=int(request.GET['id']))
                    data['planificarponencias'] = planificarponencias
                    form = PlanificarPonenciasForm(initial={'tema':planificarponencias.tema,
                                                            'fechainicio': planificarponencias.fecha_inicio,
                                                            'fechafin': planificarponencias.fecha_fin,
                                                            'justificacion': planificarponencias.justificacion,
                                                            'link': planificarponencias.link,
                                                            'nombre': planificarponencias.nombre,
                                                            'pais': planificarponencias.pais,
                                                            'sugerenciacongreso': planificarponencias.sugerenciacongreso})
                    data['form'] = form
                    return render(request, "inv_ponencia/editponencia.html", data)
                except Exception as ex:
                    pass

            elif action == 'evidenciasponenciacodigo':
                try:
                    data['title'] = u'Descargar Evidencias por Código'
                    template = get_template("inv_ponencia/evidenciasponenciascodigo.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content, 'title': data['title']})
                except Exception as ex:
                    msg = ex.__str__()
                    return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al obtener los datos. [%s]" % msg, "showSwal": "True", "swalType": "error"})

            return HttpResponseRedirect(request.path)
        else:
            data['title'] = u'Listado de Ponencias'
            search = None
            ids = None
            tipobus = None
            inscripcionid = None
            if 'id' in request.GET:
                data['tipobus'] = 2
                ids = request.GET['id']
                ponencias = PonenciasInvestigacion.objects.select_related().filter(Q(status=True) | Q(status=False, eliminadoxdoc=True), pk=int(encrypt(ids))).order_by('-fechapublicacion')
            elif 's' in request.GET:
                search = request.GET['s']
                # if search.isdigit():
                #     data['tipobus'] = 2
                #     ponencias = PonenciasInvestigacion.objects.select_related().filter(pk=search,status=True).order_by('-fechapublicacion')
                # else:
                tipobus = int(request.GET['tipobus'])
                data['tipobus'] = tipobus
                if tipobus == 2:
                    ponencias = PonenciasInvestigacion.objects.select_related().filter(Q(status=True) | Q(status=False, eliminadoxdoc=True), nombre__icontains=search).order_by('-fechapublicacion')
                if tipobus == 3:
                    if ' ' in search:
                        s = search.split(" ")
                        participantesponenciasdoc = ParticipantePonencias.objects.values_list('ponencias_id').filter(
                            status=True).filter(Q(profesor__persona__apellido1__contains=s[0]) & Q(
                            profesor__persona__apellido2__contains=s[1])).distinct()
                        ponenciasdoc = PonenciasInvestigacion.objects.select_related().filter(Q(status=True) | Q(status=False, eliminadoxdoc=True), pk__in=participantesponenciasdoc)

                        participantesponenciasadm = ParticipantePonencias.objects.values_list('ponencias_id').filter(
                            status=True).filter(Q(administrativo__persona__apellido1__contains=s[0]) & Q(
                            administrativo__persona__apellido2__contains=s[1])).distinct()
                        ponenciasadm = PonenciasInvestigacion.objects.select_related().filter(Q(status=True) | Q(status=False, eliminadoxdoc=True), pk__in=participantesponenciasadm)
                        ponencias = ponenciasdoc | ponenciasadm
                        ponencias = ponencias.order_by('-fechapublicacion')
                    else:
                        participantesponenciasdoc = ParticipantePonencias.objects.values_list('ponencias_id').filter(
                            status=True).filter(
                            Q(profesor__persona__nombres__icontains=search) |
                            Q(profesor__persona__apellido1__icontains=search) |
                            Q(profesor__persona__apellido2__icontains=search)).distinct()
                        ponenciasdoc = PonenciasInvestigacion.objects.select_related().filter(Q(status=True) | Q(status=False, eliminadoxdoc=True), pk__in=participantesponenciasdoc)

                        participantesponenciasadm = ParticipantePonencias.objects.values_list('ponencias_id').filter(
                            status=True).filter(
                            Q(administrativo__persona__nombres__icontains=search) |
                            Q(administrativo__persona__apellido1__icontains=search) |
                            Q(administrativo__persona__apellido2__icontains=search)).distinct()
                        ponenciasadm = PonenciasInvestigacion.objects.select_related().filter(Q(status=True) | Q(status=False, eliminadoxdoc=True), pk__in=participantesponenciasadm)
                        ponencias = ponenciasdoc | ponenciasadm
                        ponencias = ponencias.order_by('-fechapublicacion')
                if tipobus == 4:
                    # Buscar por nombre del congreso
                    ponencias = PonenciasInvestigacion.objects.select_related().filter(Q(status=True) | Q(status=False, eliminadoxdoc=True), evento__icontains=search).order_by('-fechapublicacion')
                if tipobus == 5:
                    # Buscar por año de publicación
                    anio = int(search)
                    ponencias = PonenciasInvestigacion.objects.select_related().filter(Q(status=True) | Q(status=False, eliminadoxdoc=True), fechapublicacion__year=anio).order_by('-fechapublicacion')

            elif 'idsp' in request.GET:
                data['tipobus'] = 3
                lista_ids_personas = request.GET['idsp'].split(",")

                participantesponencias = ParticipantePonencias.objects.values_list('ponencias_id').filter(status=True).filter(Q(profesor__persona__id__in=lista_ids_personas) |
                                                                                                                              Q(administrativo__persona__id__in=lista_ids_personas)|
                                                                                                                              Q(inscripcion__persona__id__in=lista_ids_personas)).distinct()

                ponencias = PonenciasInvestigacion.objects.select_related().filter(Q(status=True) | Q(status=False, eliminadoxdoc=True), pk__in=participantesponencias).order_by('-fechapublicacion')
            else:
                data['tipobus'] = 2
                ponencias = PonenciasInvestigacion.objects.filter(Q(status=True) | Q(status=False, eliminadoxdoc=True)).order_by('-fechapublicacion')

            estadoparticipante = 0
            if 'estadoparticipante' in request.GET:
                estadoparticipante = int(request.GET['estadoparticipante'])
                if estadoparticipante > 0:
                    if estadoparticipante == 1:
                        # participantesponencias = ParticipantePonencias.objects.values_list('ponencias_id').distinct('ponencias_id')  # .filter(status=True)#
                        # ponencias = ponencias.filter(pk__in=participantesponencias)
                        ponencias = ponencias.filter(aprobado=True)
                    else:
                        ponencias = ponencias.filter(aprobado=False)
                        # ponencias = ponencias.filter(participanteponencias__isnull=True)

            totalponencias = ponencias.count()
            totalsinparticipantes = ponencias.filter(participanteponencias__isnull=True).count()
            totalconparticipantes = totalponencias - totalsinparticipantes

            totalaprobados = ponencias.filter(aprobado=True).count()
            totalporaprobar = totalponencias - totalaprobados

            paging = MiPaginador(ponencias, 25)
            p = 1
            try:
                paginasesion = 1
                if 'paginador' in request.session:
                    paginasesion = int(request.session['paginador'])
                if 'page' in request.GET:
                    p = int(request.GET['page'])
                else:
                    p = paginasesion
                try:
                    page = paging.page(p)
                except:
                    p = 1
                page = paging.page(p)
            except:
                page = paging.page(p)
            request.session['paginador'] = p
            data['paging'] = paging
            data['page'] = page
            data['rangospaging'] = paging.rangos_paginado(p)
            data['ponencias'] = page.object_list
            data['search'] = search if search else ""
            data['ids'] = ids if ids else ""
            data['decanodircar'] = False

            data['totalponencias'] = totalponencias
            data['totalconparticipantes'] = totalconparticipantes
            data['totalsinparticipantes'] = totalsinparticipantes
            data['estadoparticipante'] = estadoparticipante
            data['totalaprobados'] = totalaprobados
            data['totalporaprobar'] = totalporaprobar
            data['aniosevidencias'] = PonenciasInvestigacion.objects.filter(status=True, fechapublicacion__isnull=False).annotate(anio=ExtractYear('fechapublicacion')).values_list('anio', flat=True).order_by('-anio').distinct()

            data['participantes'] = Persona.objects.values('id', 'cedula', 'nombres', 'apellido1', 'apellido2').filter(Q(profesor__participanteponencias__isnull=False, profesor__participanteponencias__status=True)
                                                                                                                       |Q(administrativo__participanteponencias__isnull=False, administrativo__participanteponencias__status=True)
                                                                                                                       |Q(inscripcion__participanteponencias__isnull=False, inscripcion__participanteponencias__status=True)).distinct().order_by('apellido1', 'apellido2', 'nombres')


            return render(request, "inv_ponencia/view.html", data)
