# -*- coding: UTF-8 -*-
import io
import random
import xlrd
import xlwt
from xlwt import *
from xlwt import easyxf
from openpyxl import load_workbook
import json
from itertools import chain
from django.contrib.auth.decorators import login_required
from django.db import transaction, connection
from django.db.models import Q, Max
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.shortcuts import render,redirect
from django.template.context import Context
from django.template.loader import get_template

from decorators import secure_module
from sagest.forms import PermisoInstitucionalDetalleForm, PermisoInstitucionalAdicionarForm, \
    KardexVacacionesForm, KardexVacacionesDetalleForm, KardexVacacionesIndividualForm, PermisoInstitucionalFechaForm, \
    InformesPermisoForm, SubirPermisoMasivoForm
from sagest.models import PermisoInstitucional, PermisoAprobacion, PermisoInstitucionalDetalle, TipoPermiso, \
    DistributivoPersona, TipoPermisoDetalle, RegimenLaboral, IngresoPersonal, KardexVacacionesDetalle, Departamento, \
    ESTADO_PERMISOS, SolicitudJustificacionMarcada, DetalleSolicitudJustificacionMarcada, HistorialSolicitudJustificacionMarcada, LogMarcada, \
    LogDia, RegistroMarcada, MarcadasDia
from sagest.th_marcadas import calculando_marcadas
from settings import EMAIL_DOMAIN, PUESTO_ACTIVO_ID, ARCHIVO_TIPO_GENERAL
from sga.commonviews import adduserdata
from sga.funciones import MiPaginador, log, generar_nombre, convertir_fecha, null_to_numeric, variable_valor, notificacion
from datetime import datetime,timedelta
from sga.funcionesxhtml2pdf import conviert_html_to_pdf
from sga.models import Persona, Archivo, CUENTAS_CORREOS
from sga.tasks import send_html_mail
from sga.templatetags.sga_extras import encrypt

@login_required(redirect_field_name='ret', login_url='/loginsagest')
@secure_module
@transaction.atomic()
def view(request):
    data = {}
    adduserdata(request, data)
    data['persona']=persona = request.session['persona']
    usuario = request.user
    if request.method == 'POST':
        action = request.POST['action']

        if action == 'addaprobacion':
            try:
                dias = 0
                horas = 0
                minutos = 0
                kardex = None
                total = None
                diassaldo = None
                horasaldo = None
                minutosaldo = None
                regimen = None
                i = 1
                concepto = 'Permiso '
                permiso = PermisoInstitucional.objects.get(pk=request.POST['id'])
                aprobar = PermisoAprobacion(permisoinstitucional=permiso,
                                            fechaaprobacion=datetime.now().date(),
                                            observacion=request.POST['obse'],
                                            aprueba=persona,
                                            estadosolicitud=int(request.POST['esta']))
                aprobar.save(request)
                # descuentoanterior=aprobar.permisoinstitucional.descuentovacaciones
                # tipoanterior=aprobar.permisoinstitucional.tipopermiso_id
                aprobar.permisoinstitucional.descuentovacaciones = True if int(request.POST['vaca']) == 1 else False
                aprobar.permisoinstitucional.save(request)
                permiso.actulizar_estado(request)
                if int(request.POST['esta']) == 2:
                    permiso.estadosolicitud = 4
                    permiso.save(request)
                if int(request.POST['esta']) == 1:
                    permiso.estadosolicitud = 3
                    permiso.save(request)
                if int(request.POST['esta']) == 4:
                    permiso.estadosolicitud = 6
                    permiso.save(request)
                    notificacion('Corrección de permiso institucional', ('Su permiso %s tiene observaciones ' % permiso),
                                 permiso.solicita,
                                 None, ('/th_permiso?id=%s' % permiso.id),
                                 permiso.pk,
                                 3, 'sga-sagest', PermisoInstitucional, request)
                    send_html_mail("Corrección de permiso institucional",
                                   "th_permiso_institucional/email/correo_base.html",
                                   {
                                       'receptor': permiso.solicita,
                                       'mensaje': ('Su permiso %s tiene las siguientes observaciones:' % permiso.codificacion()),
                                       'mensajeextra': aprobar.observacion,
                                   },
                                   permiso.solicita.lista_emails_envio(), [],
                                   cuenta=CUENTAS_CORREOS[0][1])
                if not request.user.has_perm('sagest.ver_permisos_salud_ocupa'):
                    aprobar.mail_notificar_talento_humano(request.session['nombresistema'])
                if permiso.estadosolicitud == 3:
                    if permiso.solicita.distributivopersona_set.exists():
                        for p in permiso.permisoinstitucionaldetalle_set.all():
                            calculando_marcadas(request, p.fechainicio, p.fechafin, permiso.solicita)
                log(u'Aprobar solicitud(TH): %s' % aprobar, request, "add")
                if aprobar.estadosolicitud == 1 and (permiso.tipopermiso.id == 3 or permiso.tipopermiso.id == 21):
                    per = Persona.objects.get(id=permiso.solicita_id)
                    permlactancia = per.maternidad()
                    permlactancia.lactancia = True
                    permlactancia.status_lactancia = True
                    permlactancia.save(request)
                # fechaprimerini=None
                # fechaprimerfin=None
                # puesto=permiso.denominacionpuesto.id
                if permiso.estadosolicitud == 3 or permiso.estadosolicitud==2 and (int(request.POST['vaca']) == 1 or permiso.descuentovacaciones):
                    if permiso.tipopermiso_id == 12:
                        concepto = concepto + 'Imputable Vacaciones'
                    elif permiso.tipopermiso_id == 24 or permiso.tipopermiso_id == 27:
                        concepto = concepto + 'Planificado'
                    else:
                        concepto = concepto + 'Descuento a Vacaciones'
                    if permiso.regimenlaboral:
                        regimen=permiso.regimenlaboral
                    else:
                        if DistributivoPersona.objects.filter(persona=permiso.solicita,denominacionpuesto=permiso.denominacionpuesto).exists():
                            regimen=DistributivoPersona.objects.filter(persona=permiso.solicita,denominacionpuesto=permiso.denominacionpuesto)[0].regimenlaboral
                    if regimen:
                        if KardexVacacionesDetalle.objects.filter(kardex__persona=permiso.solicita,kardex__regimenlaboral=regimen).exists():
                            if permiso.estadosolicitud == 3  and (int(request.POST['vaca']) == 1 or permiso.descuentovacaciones or permiso.tipopermiso_id == 24 or permiso.tipopermiso_id == 27 or permiso.tipopermiso_id == 12):
                                dias=int(request.POST['dias'])
                                horas=int(request.POST['horas'])
                                minutos=int(request.POST['minutos'])
                                # detalle = permiso.permisoinstitucionaldetalle_set.filter(status=True)
                                # if permiso.tipopermiso_id == 24 or permiso.tipopermiso_id == 27 or (descuentoanterior == False and tipoanterior!=12):
                                #     for d in detalle:
                                #         if i == 1:
                                #             inicio = timedelta(hours=d.horainicio.hour, minutes=d.horainicio.minute)
                                #             fin = timedelta(hours=d.horafin.hour, minutes=d.horafin.minute)
                                #             total = fin - inicio
                                #             fechaprimerini = d.fechainicio
                                #             fechaprimerfin = d.fechafin
                                #             # horas = d.horafin.hour - d.horainicio.hour
                                #             # minutos = d.horafin.minute - d.horainicio.minute
                                #             horas = total.seconds // 3600
                                #             minutos = (total.seconds // 60) - (horas * 60)
                                #             if d.fechafin != d.fechainicio:
                                #                 dias = d.fechafin - d.fechainicio
                                #                 dias = dias.days
                                #             if horas >= 8:
                                #                 dias = dias + 1
                                #                 horas = 0
                                #             i = i + 1
                                #         else:
                                #             # horas = horas + (d.horafin.hour - d.horainicio.hour)
                                #             # minutos = minutos + (d.horafin.minute - d.horainicio.minute)
                                #             inicio = timedelta(hours=d.horainicio.hour, minutes=d.horainicio.minute)
                                #             fin = timedelta(hours=d.horafin.hour, minutes=d.horafin.minute)
                                #             total1 = fin - inicio
                                #             horas = horas + (total1.seconds // 3600)
                                #             minutos = minutos + ((total1.seconds // 60) - ((total1.seconds // 3600) * 60))
                                #             if fechaprimerini != d.fechainicio and fechaprimerfin != d.fechafin:
                                #                 dias = dias + (d.fechafin - d.fechainicio)
                                #                 dias = dias.days
                                #             if horas >= 8:
                                #                 dias = dias + 1
                                #                 horas = 0
                                # elif permiso.tipopermiso_id == 12 and descuentoanterior == True:
                                #     for d in detalle:
                                #         if i==1:
                                #             inicio = timedelta(hours=d.horainicio.hour, minutes=d.horainicio.minute)
                                #             fin = timedelta(hours=d.horafin.hour, minutes=d.horafin.minute)
                                #             total = fin - inicio
                                #             horas = total.seconds // 3600
                                #             minutos = (total.seconds // 60) - (horas * 60)
                                #             # horas = d.horafin.hour - d.horainicio.hour
                                #             # minutos = d.horafin.minute - d.horainicio.minute
                                #             i = i + 1
                                #         else:
                                #             inicio = timedelta(hours=d.horainicio.hour, minutes=d.horainicio.minute)
                                #             fin = timedelta(hours=d.horafin.hour, minutes=d.horafin.minute)
                                #             total1 = fin - inicio
                                #             horas = horas + (total1.seconds // 3600)
                                #             minutos = minutos + ((total1.seconds // 60) - ((total1.seconds // 3600) * 60))
                                #             # horas =horas+ (d.horafin.hour - d.horainicio.hour)
                                #             # minutos =minutos+(d.horafin.minute - d.horainicio.minute)
                                #     dias=0
                                #     if horas>=8:
                                #         dias=1
                                #         horas=0

                                if IngresoPersonal.objects.filter(persona=permiso.solicita, regimenlaboral=regimen).exists():
                                    kardex=IngresoPersonal.objects.filter(status=True,persona=permiso.solicita, estado=1, regimenlaboral=regimen).order_by("-id")[0]
                                    detallekardex=kardex.kardexvacacionesdetalle_set.filter(kardex__persona=permiso.solicita,kardex__regimenlaboral=regimen,status=True).latest('id')
                                    # detallekardex=KardexVacacionesDetalle.objects.filter(kardex__persona=permiso.solicita,kardex__regimenlaboral=regimen).latest('id')
                                    if detallekardex:
                                        saldo = timedelta(days=detallekardex.diasal, hours=detallekardex.horasal,minutes=detallekardex.minsal)
                                    else:
                                        saldo = timedelta(days=0, hours=0,minutes=0)

                                    permisos = timedelta(days=dias, hours=horas, minutes=minutos)
                                    if detallekardex.diasal<1 and detallekardex.horasal<1 and detallekardex.minsal<1:
                                        transaction.set_rollback(True)
                                        return JsonResponse({"result": "bad", "mensaje": u"Lo sentimos ya no tiene saldo."})
                                    elif saldo<permisos:
                                        transaction.set_rollback(True)
                                        return JsonResponse({"result": "bad", "mensaje": u"Lo sentimos ya no tiene saldo."})
                                    total=saldo-permisos
                                    diassaldo=total.days
                                    horasaldo = total.seconds // 3600
                                    minutosaldo= (total.seconds // 60) - (horasaldo * 60)
                                    detalle=KardexVacacionesDetalle(kardex=kardex,
                                                                    permiso=permiso,
                                                                    fecha=datetime.now(),
                                                                    operacion=2,
                                                                    concepto=concepto,
                                                                    diava=dias,horava=horas,minva=minutos,
                                                                    diasal=diassaldo,horasal=horasaldo,minsal=minutosaldo)
                                    detalle.save(request)
                                else:
                                    transaction.set_rollback(True)
                                    return JsonResponse({"result": "bad", "mensaje": u"No tiene cardex, para su régimen"})
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": "bad", "mensaje": u"No tiene régimen."})
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        if action == 'addaprobacion_rechazar':
            try:
                dias = 0
                horas = 0
                minutos = 0
                kardex = None
                total = None
                diassaldo = None
                horasaldo = None
                minutosaldo = None
                i = 1
                concepto = 'Permiso '
                permiso = PermisoInstitucional.objects.get(pk=request.POST['id'])
                aprobar = PermisoAprobacion(permisoinstitucional=permiso,
                                            fechaaprobacion=datetime.now().date(),
                                            observacion=request.POST['obse'],
                                            aprueba=persona,
                                            estadosolicitud=int(request.POST['esta']))
                aprobar.save(request)
                descuentoanterior=aprobar.permisoinstitucional.descuentovacaciones
                tipoanterior=aprobar.permisoinstitucional.tipopermiso_id
                aprobar.permisoinstitucional.descuentovacaciones = False
                aprobar.permisoinstitucional.save(request)
                permiso.estadosolicitud=4
                permiso.save(request)
                kardex = permiso.kardexvacacionesdetalle_set.filter(status=True)
                if kardex:
                    kardex=kardex[0]
                    ingreso = kardex.kardex
                    saldo = timedelta(days=kardex.diasal, hours=kardex.horasal,
                                      minutes=kardex.minsal)
                    permisos = timedelta(days=kardex.diava, hours=kardex.horava, minutes=kardex.minva)
                    total = saldo + permisos
                    diassaldo = total.days
                    horasaldo = total.seconds // 3600
                    minutosaldo = (total.seconds // 60) - (horasaldo * 60)

                    horas = permisos.seconds // 3600
                    minutos = (permisos.seconds // 60) - (horas * 60)

                    detalle = KardexVacacionesDetalle(kardex=ingreso,
                                                          fecha=datetime.now(),
                                                          operacion=1,
                                                          concepto="Devolución por rechazo de permiso "+ str(permiso) ,
                                                          diava=permisos.days, horava=horas, minva=minutos,
                                                          diasal=diassaldo, horasal=horasaldo, minsal=minutosaldo)
                    detalle.save(request)

                aprobar.mail_notificar_talento_humano(request.session['nombresistema'])
                if permiso.estadosolicitud == 3:
                    if permiso.solicita.distributivopersona_set.exists():
                        for p in permiso.permisoinstitucionaldetalle_set.all():
                            calculando_marcadas(request, p.fechainicio, p.fechafin, permiso.solicita)
                log(u'Aprobar solicitud(TH): %s' % aprobar, request, "add")
                if permiso.tipopermiso.status==True and permiso.tipopermiso_id == 12 :
                    concepto=concepto+'Imputable Vacaciones'
                else:
                    concepto = concepto + 'Planificado'
                # if KardexVacacionesDetalle.objects.filter(permiso=permiso,status=True).exists():
                #     KardexVacacionesDetalle.objects.filter(permiso=permiso, status=True).update(status=False)
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'listapermisos':
            try:
                regimenlaboralpersona = extraer_regimen_laboral_antes_guardar(int(request.POST['id']))
                if regimenlaboralpersona:
                    tipospermiso = TipoPermiso.objects.filter(
                        tipopermisoregimenlaboral__regimenlaboral=regimenlaboralpersona, status=True,
                        tipopermisodetalle__tipopermiso__isnull=False, tipopermisodetalle__vigente=True).distinct()
                    return JsonResponse({"result": "ok", "tipospermiso": [
                        {"id": tipopermiso.id, "descripcion": tipopermiso.descripcion} for tipopermiso in
                        tipospermiso]})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al consultar los datos."})

        elif action == 'listapuesto':
            try:
                puesto=DistributivoPersona.objects.filter(id=int(request.POST['id']), status=True,estadopuesto__id=PUESTO_ACTIVO_ID).distinct()
                return JsonResponse({"result": "ok", "tipospermiso": [{"id": puest.id, "descripcion": puest.denominacionpuesto.descripcion,"regimen":puest.regimenlaboral.descripcion} for puest in puesto],"count":puesto.__len__()})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al consultar los datos."})

        elif action == 'add':
            try:
                f = PermisoInstitucionalAdicionarForm(request.POST, request.FILES)
                newfile = None
                if 'archivo' in request.FILES:
                    newfile = request.FILES['archivo']
                    if newfile:
                        newfilesd = newfile._name
                        ext = newfilesd[newfilesd.rfind("."):]
                        if not ext == '.pdf':
                            return JsonResponse({"result": "bad", "mensaje": u"Error, solo archivos .pdf."})
                        if newfile.size > 10485760:
                            return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 10 Mb."})
                        if newfile:
                            newfile._name = generar_nombre("Solicitud", newfile._name)
                else:
                    if TipoPermisoDetalle.objects.get(pk=int(request.POST['tipopermisodetalle'])).perdirarchivo:
                        return JsonResponse({"result": "bad", "mensaje": u"Error, debe subir un archivo."})

                if f.is_valid():
                    if f.cleaned_data['persona'] < 1:
                        return JsonResponse({"result": "bad", "mensaje": u"Seleccione Persona."})
                    fecha = datetime.now().date()
                    datos = json.loads(request.POST['lista_items1'])
                    tipopermisodetalle = f.cleaned_data['tipopermisodetalle']
                    distributivo = f.cleaned_data['denominacionpuesto']
                    # distributivo = DistributivoPersona.objects.filter(id=int(f.cleaned_data['persona']),denominacionpuesto=denominacionpuesto.denominacionpuesto)[0]
                    personaadmin = distributivo.persona
                    lista = []
                    if not datos:
                        return JsonResponse({"result": "bad", "mensaje": u"Debe ingresar la duración del permiso."})
                    for d in datos:
                        fechainicio = convertir_fecha(d['fechainicio'])
                        horainicio = int(d['horainicio'].split(':')[0])
                        minutoinicio = int(d['horainicio'].split(':')[1])
                        fechafin = convertir_fecha(d['fechafin'])
                        horafin = int(d['horafin'].split(':')[0])
                        minutofin = int(d['horafin'].split(':')[1])
                        resultado_en_dia = valida_este_en_dias(tipopermisodetalle, fechainicio, fechafin)
                        if not resultado_en_dia[0]:
                            return JsonResponse({"result": "bad", "mensaje": resultado_en_dia[1]})
                        lista.append([datetime(fechainicio.year, fechainicio.month, fechainicio.day, horainicio, minutoinicio), datetime(fechafin.year, fechafin.month, fechafin.day, horafin, minutofin)])
                    p = 1

                    def hora_dia(hr, mi):
                        now = datetime.now()
                        return now.replace(hour=hr, minute=mi, second=0, microsecond=0)

                    for elemento in lista:
                        otros = lista
                        otros.remove(elemento)
                        for otro in otros:
                            if otro[0] <= elemento[0] <= otro[1]:
                                if hora_dia(otro[0].hour, otro[0].minute) <= hora_dia(elemento[0].hour, elemento[0].minute) <= hora_dia(otro[1].hour, otro[1].minute):
                                    return JsonResponse({"result": "bad", "mensaje": u"Error fechas incorrectas."})
                            if otro[0] <= elemento[1] <= otro[1]:
                                if hora_dia(otro[0].hour, otro[0].minute) <= hora_dia(elemento[1].hour, elemento[1].minute) <= hora_dia(otro[1].hour, otro[1].minute):
                                    return JsonResponse({"result": "bad", "mensaje": u"Error fechas incorrectas."})
                        p += 1
                    secu = null_to_numeric(PermisoInstitucional.objects.filter(solicita=personaadmin, fechasolicitud__year=fecha.year).aggregate(secu=Max("secuencia"))['secu'])
                    secuencia = secu + 1
                    if not personaadmin.mis_plantillas_actuales().filter(denominacionpuesto=distributivo.denominacionpuesto).exists():
                        return JsonResponse({"result": "bad", "mensaje": u"Error no existe la unidad organica."})
                    unidadorganica = personaadmin.mis_plantillas_actuales().filter(denominacionpuesto=distributivo.denominacionpuesto)[0].unidadorganica
                    if not unidadorganica.responsable:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": "bad", "mensaje": u"No existe un responsable del departamento, consulte con talento humano."})

                    if 'permisofamilia' in request.POST:
                        permisofamilia = f.cleaned_data['permisofamilia']
                    else:
                        permisofamilia = None
                    permiso = PermisoInstitucional(fechasolicitud=fecha,
                                                   solicita=personaadmin,
                                                   secuencia=secuencia,
                                                   permisofamilia=permisofamilia,
                                                   tiposolicitud=1,
                                                   tipopermiso=f.cleaned_data['tipopermiso'],
                                                   tipopermisodetalle=f.cleaned_data['tipopermisodetalle'],
                                                   descuentovacaciones=False if not tipopermisodetalle else tipopermisodetalle.descuentovacaciones,
                                                   motivo=f.cleaned_data['motivo'],
                                                   estadosolicitud=1,
                                                   denominacionpuesto=distributivo.denominacionpuesto,
                                                   regimenlaboral=distributivo.regimenlaboral,
                                                   unidadorganica=unidadorganica,
                                                   casasalud=f.cleaned_data['casasalud'])
                    permiso.save(request)
                    if newfile:
                        permiso.archivo = newfile
                        permiso.save(request)
                    for d in datos:
                        detalleingprod = PermisoInstitucionalDetalle(permisoinstitucional=permiso,
                                                                     fechainicio=convertir_fecha(d['fechainicio']),
                                                                     fechafin=convertir_fecha(d['fechafin']),
                                                                     horainicio=d['horainicio'],
                                                                     horafin=d['horafin'])
                        detalleingprod.save(request)
                        detalleingprod.usuario_creacion_id = personaadmin.usuario.id
                        detalleingprod.save(request)
                    permiso.usuario_creacion_id=personaadmin.usuario.id
                    permiso.save(request)
                    # send_html_mail("Solicitud de permiso[%s]" % permiso.solicita, "emails/permisosolicita.html", {'sistema': request.session['nombresistema'], 'codificacion': permiso.codificacion(), 'responsable': permiso.unidadorganica.responsable, 'solicita': permiso.solicita, 't': miinstitucion()}, permiso.unidadorganica.responsable.lista_emails(), [], cuenta=CUENTAS_CORREOS[1][1])
                    log(u'Adiciono nueva solicitud de permiso: %s' % permiso, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'listarregimen':
            try:
                id = request.POST['id']
                regimen = DistributivoPersona.objects.filter(status=True, persona__id=id, regimenlaboral__status=True,estadopuesto__id=PUESTO_ACTIVO_ID).distinct()
                return JsonResponse({"result": "ok", "regimen": [{"id": reg.regimenlaboral.id, "descripcion": reg.regimenlaboral.descripcion} for reg in regimen],"count": regimen.__len__()
                                     })
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al consultar los datos."})

        elif action == 'actualizar':
            try:
                ingreso=IngresoPersonal.objects.get(pk=int(request.POST['id']))
                if ingreso.fechasalida:
                    ingreso.estado=2
                    ingreso.save(request)
                else:
                    return JsonResponse({"result": "bad", "mensaje": u"Esta activo"})
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al renovar."})

        elif action == 'addkardexindividualdetalle':
            try:
                dias=None
                horas=None
                minutos=None
                f = KardexVacacionesDetalleForm(request.POST, request.FILES)
                if f.is_valid():
                    kardex = IngresoPersonal.objects.get(id=int(request.POST['id']))
                    regimen = kardex.regimenlaboral
                    persona = kardex.persona
                    dias = int(f.cleaned_data['diava'])
                    horas = int(f.cleaned_data['horava'])
                    minutos = int(f.cleaned_data['minva'])
                    operacion = int(f.cleaned_data['operacion'])
                    permisos = timedelta(days=dias, hours=horas, minutes=minutos)
                    # kardex = IngresoPersonal.objects.filter(status=True, persona=persona, estado=1, regimenlaboral__id=regimen).order_by("-id")[0]
                    if kardex.kardexvacacionesdetalle_set.filter(status=True).exists():
                        detallekardex = kardex.kardexvacacionesdetalle_set.filter(status=True).latest('id')
                        saldo = timedelta(days=detallekardex.diasal, hours=detallekardex.horasal,minutes=detallekardex.minsal)
                    else:
                        saldo = timedelta(days=0, hours=0, minutes=0)
                    if operacion == 2:
                        if detallekardex.diasal < 1 and detallekardex.horasal < 1 and detallekardex.minsal < 1:
                            return JsonResponse({"result": "bad", "mensaje": u"Lo sentimos ya no tiene saldo, tiene: 0d 0h 0m"})
                        elif saldo < permisos:
                            return JsonResponse({"result": "bad", "mensaje": u"Lo sentimos ya no tiene saldo, tiene: "+str(detallekardex.diasal)+"d "+str(detallekardex.horasal)+"h "+str(detallekardex.minsal)+"m"})
                    if operacion == 1:
                        total = saldo + permisos
                    elif operacion == 2 :
                        total = saldo - permisos
                    diassaldo = total.days
                    horasaldo = total.seconds // 3600
                    minutosaldo = (total.seconds // 60) - (horasaldo * 60)
                    permiso=f.cleaned_data['permiso']
                    if permiso:
                        detalle = KardexVacacionesDetalle(kardex=kardex,
                                                          permiso=permiso,
                                                          fecha=datetime.now(),
                                                          operacion=operacion,
                                                          concepto=f.cleaned_data['concepto'],
                                                          diava=dias, horava=horas, minva=minutos,
                                                          diasal=diassaldo, horasal=horasaldo, minsal=minutosaldo)
                    else:
                        detalle = KardexVacacionesDetalle(kardex=kardex,
                                                          fecha=datetime.now(),
                                                          operacion=operacion,
                                                          concepto=f.cleaned_data['concepto'],
                                                          diava=dias, horava=horas, minva=minutos,
                                                          diasal=diassaldo, horasal=horasaldo, minsal=minutosaldo)
                    detalle.save(request)
                    kardex.recalcular()
                    return JsonResponse({"result": False}, safe=False)
                return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()], "message": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'validarfechadiasplazo':
            try:
                ne = request.POST['ne']
                if not ne[:1] == 'e':
                    detalle = TipoPermisoDetalle.objects.get(pk=int(request.POST['id']))
                    mensaje = ''
                    resultado_valida_dia = valida_este_en_dias(detalle, convertir_fecha(request.POST['fd']), convertir_fecha(request.POST['ff']))
                    puedeadicionar = resultado_valida_dia[0]
                    if not resultado_valida_dia[0]:
                        mensaje = resultado_valida_dia[1]
                    return JsonResponse({"result": "ok", 'puedeadicionar':puedeadicionar, "mensaje":mensaje})
                else:
                    return JsonResponse({"result": "ok", 'puedeadicionar': True})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al elminar los datos."})

        elif action == 'reportepermisosrechazadopdf':
            try:
                listado = []
                fechainicio = datetime.strptime(request.POST['ini'], "%Y-%m-%d").date()
                fechafin = datetime.strptime(request.POST['fin'], "%Y-%m-%d").date()
                if fechafin <= fechainicio:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al descargar reporte."})

                if fechafin >= datetime.now().date():
                    fechafin = datetime.now().date() + timedelta(days=-1)
                iddepartamentos_sinordenar = PermisoInstitucional.objects.values_list('unidadorganica_id').filter(fechasolicitud__range=(fechainicio, fechafin)).distinct('unidadorganica_id').order_by('unidadorganica_id')
                iddepartamentos = Departamento.objects.values_list('id').filter(pk__in=iddepartamentos_sinordenar).order_by('nombre')
                for iddepartamento in iddepartamentos:
                    departamento = Departamento.objects.get(pk=iddepartamento[0])
                    pendiente = PermisoInstitucional.objects.values('id').filter(estadosolicitud=1, unidadorganica_id=iddepartamento[0], fechasolicitud__range=(fechainicio, fechafin)).count()
                    solicitada = PermisoInstitucional.objects.values('id').filter(unidadorganica_id=iddepartamento[0], fechasolicitud__range=(fechainicio, fechafin)).count()
                    contaraprobadodirector = 0
                    contaraprobadouath = 0
                    permisoaprobados = PermisoInstitucional.objects.filter((Q(estadosolicitud=3) | Q(estadosolicitud=4)), unidadorganica_id=iddepartamento[0], fechasolicitud__range=(fechainicio, fechafin))
                    for permisoaprobado in permisoaprobados:
                        if permisoaprobado.contar_aprobados() == 1:
                            contaraprobadodirector += 1
                        else:
                            contaraprobadodirector += 1
                            contaraprobadouath += 1
                    listado.append([departamento.nombre, solicitada, pendiente, contaraprobadodirector, contaraprobadouath])
                return conviert_html_to_pdf('th_permiso_institucional/reportepermisosaprobado.html',{
                    'pagesize': 'A4',
                    'listado': listado,
                    'hoy': datetime.now(),
                    'fechainicio': fechainicio,
                    'fechafin': fechafin})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al descargar reporte."})
                pass

        elif action == 'addkardexindividual':
            try:
                f = KardexVacacionesIndividualForm(request.POST, request.FILES)
                if f.is_valid():
                    if f.cleaned_data['persona'] < 1:
                        return JsonResponse({"result": "bad", "mensaje": u"Seleccione Persona."})
                    persona=Persona.objects.get(pk=int(f.cleaned_data['persona']))
                    regimenlaboral = f.cleaned_data['regimenlaboral']
                    fechaingreso = f.cleaned_data['fechaingreso']
                    fechasalida = f.cleaned_data['fechasalida']
                    estado = f.cleaned_data['estado']
                    nombramiento = f.cleaned_data['nombramiento']
                    contratoindefinido = f.cleaned_data['contratoindefinido']
                    kardex = IngresoPersonal(persona=persona,
                                             regimenlaboral=regimenlaboral,
                                             fechaingreso=fechaingreso,
                                             fechasalida=fechasalida,
                                             estado=estado,
                                             nombramiento=nombramiento,
                                             contratoindefinido=contratoindefinido
                                             )
                    kardex.save(request)
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'addkardex':
            try:
                form = KardexVacacionesForm(request.POST, request.FILES)
                if form.is_valid():
                    hoy = datetime.now().date()
                    nfile = request.FILES['archivo']
                    nfile._name = generar_nombre("importacion_kardex_", nfile._name)
                    archivo = Archivo(nombre='IMPORTACION KARDEX VACACIONES',
                                      fecha=datetime.now().date(),
                                      archivo=nfile,
                                      tipo_id=ARCHIVO_TIPO_GENERAL)
                    archivo.save(request)
                    workbook = xlrd.open_workbook(archivo.archivo.file.name)
                    sheet = workbook.sheet_by_index(0)
                    linea = 1
                    for rowx in range(sheet.nrows):
                        if linea >= 2:
                            cols = sheet.row_values(rowx)
                            cedula=str(cols[0]).strip()
                            cedula=cedula.replace(".0","")
                            # if not IngresoPersonal.objects.filter(persona__cedula=cedula).exists():
                            persona=Persona.objects.get(status=True,cedula=cedula)
                            fechaingreso = convertir_fecha(str(cols[1]))
                            # if DistributivoPersona.objects.filter(status=True,persona=persona).exists():
                            if str(cols[2]):
                                fechasalida = convertir_fecha(str(cols[2]))
                                kardex = IngresoPersonal(persona=persona,
                                                         regimenlaboral_id=4,
                                                         fechaingreso=fechaingreso,
                                                         fechasalida=fechasalida,
                                                         estado=2)
                            else:
                                kardex = IngresoPersonal(persona=persona,
                                                         regimenlaboral_id=4,
                                                         fechaingreso=fechaingreso,
                                                         estado=1)
                            kardex.save(request)
                            detallekardex=KardexVacacionesDetalle(
                                kardex=kardex,
                                fecha=datetime.now(),
                                operacion=1,
                                concepto='Renovacion Primera',
                                diava=0,
                                horava =0,
                                minva =0,
                                diasal = int(cols[3]),
                                horasal = int(cols[4]),
                                minsal = int(cols[5]))
                            detallekardex.save(request)
                        else:
                            linea += 1
                    log(u'Agrego un Ingreso de personal: %s' % kardex, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'delekardex':
            try:
                ingreso = IngresoPersonal.objects.get(pk=int(request.POST['id']))
                ingreso.delete()
                log(u'Eliminó un Ingreso de personal: %s' % ingreso, request, "del")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'editkardex':
            try:
                ingreso = IngresoPersonal.objects.get(pk=request.POST['id'])
                f = KardexVacacionesIndividualForm(request.POST)
                if f.is_valid():
                    if  f.cleaned_data['fechasalida']:
                        ingreso.fechaingreso = f.cleaned_data['fechaingreso']
                        ingreso.fechasalida = f.cleaned_data['fechasalida']
                        ingreso.estado = 2
                        ingreso.nombramiento=f.cleaned_data['nombramiento']
                        ingreso.contratoindefinido = f.cleaned_data['contratoindefinido']
                    else:
                        ingreso.fechaingreso = f.cleaned_data['fechaingreso']
                        ingreso.estado = f.cleaned_data['estado']
                        ingreso.nombramiento = f.cleaned_data['nombramiento']
                        ingreso.contratoindefinido = f.cleaned_data['contratoindefinido']
                    ingreso.save(request)
                    log(u'Edito un Ingreso de personal: %s' % ingreso, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'renovar':
            try:
                ingreso=IngresoPersonal.objects.get(pk=int(request.POST['id']))
                if not KardexVacacionesDetalle.objects.filter(status=True,kardex=ingreso,periodoinicio=datetime.now().year -1 ,periodofin=datetime.now().year).exists():
                    if ingreso.fechaingreso == datetime.now().date():
                        detalle=KardexVacacionesDetalle.objects.values_list("diasal","horasal","minsal").filter(status=True, kardex=ingreso, periodoinicio__isnull=True, periodofin__isnull=True)[0]
                        dias=30+detalle[0]
                        detalle = KardexVacacionesDetalle(kardex=ingreso,
                                                          fecha=datetime.now(),
                                                          operacion=1,
                                                          concepto=u"Renovación Vacaciones",
                                                          diava=0, horava=0, minva=0,
                                                          diasal=dias, horasal=detalle[1], minsal=detalle[2],
                                                          periodoinicio=datetime.now().year -1 ,
                                                          periodofin=datetime.now().year
                                                          )
                        detalle.save(request)
                        return JsonResponse({"result": "ok"})
                else:
                    return JsonResponse({"result": "bad", "mensaje": u"Ya existe renovación para este periodo."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al renovar."})

        elif action == 'permisodetalle':
            try:
                permisodetalle = TipoPermiso.objects.get(pk=request.POST['id'])
                si = 2
                if 'ENFERMEDAD' in permisodetalle.descripcion:
                    si=1
                lista = []
                for dettalle in permisodetalle.tipopermisodetalle_set.filter(status=True, vigente=True):
                    lista.append([dettalle.id,dettalle.__str__(), 1 if dettalle.perdirarchivo else 0])
                return JsonResponse({'result': 'ok', 'lista': lista, 'count': len(lista),'si':si})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'editfecha':
            try:
                f = PermisoInstitucionalFechaForm(request.POST, request.FILES)
                newfile = None
                if 'archivo' in request.FILES:
                    newfile = request.FILES['archivo']
                    if newfile.size > 4194304:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Error, Tamaño de archivo Maximo permitido es de 4Mb"})

                permiso = PermisoInstitucional.objects.get(pk=int(request.POST['id']))

                if f.is_valid():
                    datos = json.loads(request.POST['lista_items1'])
                    lista = []
                    listaeditar = []
                    if not datos:
                        return JsonResponse({"result": "bad", "mensaje": u"Debe seleccionar un rango de tiempo."})
                    for d in datos:
                        fechainicio = convertir_fecha(d['fechainicio'])
                        horainicio = int(d['horainicio'].split(':')[0])
                        minutoinicio = int(d['horainicio'].split(':')[1])
                        fechafin = convertir_fecha(d['fechafin'])
                        horafin = int(d['horafin'].split(':')[0])
                        minutofin = int(d['horafin'].split(':')[1])
                        # ne = d['ne']
                        # if not ne[:1] == 'e':
                        #     resultado = valida_este_en_fecha(tipopermisodetalle, fechainicio, permiso.fechasolicitud)
                        #     resultado_en_dia = valida_este_en_dias(tipopermisodetalle, fechainicio, fechafin)
                        #     if not resultado[0] and not resultado_en_dia[0]:
                        #         return JsonResponse(
                        #             {"result": "bad", "mensaje": resultado[1] + " Y " + resultado_en_dia[1]})
                        #     else:
                        #         if not resultado[0]:
                        #             return JsonResponse({"result": "bad", "mensaje": resultado[1]})
                        #         if not resultado_en_dia[0]:
                        #             return JsonResponse({"result": "bad", "mensaje": resultado_en_dia[1]})
                        # else:
                        #     listaeditar.append(int(ne[1:ne.__len__()]))
                        lista.append([datetime(fechainicio.year, fechainicio.month, fechainicio.day, horainicio, minutoinicio), datetime(fechafin.year, fechafin.month, fechafin.day, horafin, minutofin)])
                    p = 1

                    def hora_dia(hr, mi):
                        now = datetime.now()
                        return now.replace(hour=hr, minute=mi, second=0, microsecond=0)

                    # for elemento in lista:
                    #     otros = lista
                    #     otros.remove(elemento)
                    #     for otro in otros:
                    #         if otro[0] <= elemento[0] <= otro[1]:
                    #             if hora_dia(otro[0].hour, otro[0].minute) <= hora_dia(elemento[0].hour, elemento[0].minute) <= hora_dia(otro[1].hour, otro[1].minute):
                    #                 return JsonResponse({"result": "bad", "mensaje": u"Error fechas incorrectas."})
                    #         if otro[0] <= elemento[1] <= otro[1]:
                    #             if hora_dia(otro[0].hour, otro[0].minute) <= hora_dia(elemento[1].hour, elemento[1].minute) <= hora_dia(otro[1].hour, otro[1].minute):
                    #                 return JsonResponse({"result": "bad", "mensaje": u"Error fechas incorrectas."})
                    #     p += 1
                    # if not persona.mis_plantillas_actuales().filter(denominacionpuesto=denominacionpuesto.denominacionpuesto).exists():
                    #     return JsonResponse({"result": "bad", "mensaje": u"Error no existe la unidad organica."})
                    # if 'permisofamilia' in request.POST:
                    #     permisofamilia = f.cleaned_data['permisofamilia']
                    # else:
                    #     permisofamilia = None
                    # unidadorganica = persona.mis_plantillas_actuales().filter(denominacionpuesto=denominacionpuesto.denominacionpuesto)[0].unidadorganica
                    # permiso.tiposolicitud = f.cleaned_data['tiposolicitud']
                    # permiso.tipopermiso = f.cleaned_data['tipopermiso']
                    # permiso.tipopermisodetalle = f.cleaned_data['tipopermisodetalle']
                    # permiso.descuentovacaciones = False if not tipopermisodetalle else tipopermisodetalle.descuentovacaciones
                    # permiso.motivo = f.cleaned_data['motivo']
                    # permiso.estadosolicitud = 1
                    cuenta = DistributivoPersona.objects.filter(persona=permiso.solicita, status=True, estadopuesto__id=PUESTO_ACTIVO_ID).count()
                    dis = DistributivoPersona.objects.filter(persona=permiso.solicita, status=True, estadopuesto__id=PUESTO_ACTIVO_ID)[0]
                    if cuenta > 1:
                        permiso.denominacionpuesto = f.cleaned_data['denominacionpuesto']
                    else:
                        permiso.denominacionpuesto = dis.denominacionpuesto
                    permiso.unidadorganica = dis.unidadorganica
                    # permiso.permisofamilia = permisofamilia
                    if newfile:
                        newfile._name = generar_nombre("Solicitud", newfile._name)
                        permiso.archivo = newfile
                    permiso.save(request)
                    # if newfile:
                    #     permiso.archivo = newfile
                    #     permiso.save(request)
                    permiso.permisoinstitucionaldetalle_set.all().exclude(pk__in=listaeditar).delete()
                    for d in datos:
                        # ne = d['ne']
                        # if ne[:1]=='n':
                        detalleingprod = PermisoInstitucionalDetalle(permisoinstitucional=permiso,
                                                                     fechainicio=convertir_fecha(d['fechainicio']),
                                                                     fechafin=convertir_fecha(d['fechafin']),
                                                                     horainicio=d['horainicio'],
                                                                     horafin=d['horafin'])
                        # else:
                        #     detalleingprod = PermisoInstitucionalDetalle.objects.get(pk=int(ne[1:ne.__len__()]))
                        #     detalleingprod.horainicio = d['horainicio']
                        #     detalleingprod.horafin = d['horafin']
                        detalleingprod.save(request)
                    log(u'Modifico solicitud de permiso: %s' % permiso, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                     raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'subirdocumentosoporte':
            try:
                f = InformesPermisoForm(request.POST, request.FILES)
                if f.is_valid():
                    if 'archivo' not in request.FILES:
                        return JsonResponse({'result': 'bad', 'mensaje': u"Debe seleccionar un archivo."})
                    else:
                        newfile = request.FILES['archivo']
                        if newfile.size > 4194304:
                            return JsonResponse(
                                {"result": "bad", "mensaje": u"Error, Tamaño de archivo Maximo permitido es de 4Mb"})
                    newfile._name = generar_nombre("Solicitud", newfile._name)
                    informefac = PermisoInstitucional.objects.get(id=int(encrypt(request.POST['id'])))
                    informefac.archivo = newfile
                    informefac.save(request)
                    log(u'Subio documento soporte: %s' % informefac, request, "add")
                    return JsonResponse({"result": False, "mensaje": "Archivo subido correctamente."})
                return JsonResponse({'result': True, "mensaje": u"Error al subir el archivo."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, "mensaje": u"Error al subir el archivo."})

        elif action == 'delkardex':
            try:
                if KardexVacacionesDetalle.objects.filter(pk=request.POST['id'], status=True):
                    kardex = KardexVacacionesDetalle.objects.get(pk=request.POST['id'], status=True)
                    kardex.delete()
                    log(u'Elimino kardex vacaciones: %s' % kardex, request, "del")
                res_json = {"error": False}
            except Exception as ex:
                transaction.set_rollback(True)
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        elif action == 'updatedia':
            try:
                detalle = KardexVacacionesDetalle.objects.get(pk=int(request.POST['id']))
                valor = int(request.POST['vc'])
                detalle.diasal = valor
                detalle.save(request)
                return JsonResponse({'result': 'ok', 'valor': detalle.diasal})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', "mensaje": u"Error al actualizar dias."})

        elif action == 'updatehora':
            try:
                detalle = KardexVacacionesDetalle.objects.get(pk=int(request.POST['id']))
                valor = int(request.POST['vc'])
                detalle.horasal = valor
                detalle.save(request)
                return JsonResponse({'result': 'ok', 'valor': detalle.horasal})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', "mensaje": u"Error al actualizar horas."})
        elif action == 'updatemin':
            try:
                detalle = KardexVacacionesDetalle.objects.get(pk=int(request.POST['id']))
                valor = int(request.POST['vc'])
                detalle.minsal = valor
                detalle.save(request)
                return JsonResponse({'result': 'ok', 'valor': detalle.minsal})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', "mensaje": u"Error al actualizar minutos."})

        if action == 'addsubirinforme':
            try:
                f = InformesPermisoForm(request.POST, request.FILES)
                # 2.5MB - 2621440
                # 5MB - 5242880
                # 10MB - 10485760
                # 20MB - 20971520
                # 50MB - 5242880
                # 100MB 104857600
                # 250MB - 214958080
                # 500MB - 429916160
                d = request.FILES['archivo']
                newfilesd = d._name
                ext = newfilesd[newfilesd.rfind("."):]
                if ext == '.pdf':
                    a = 1
                else:
                    return JsonResponse({"result": "bad", "mensaje": u"Error, solo archivos .pdf."})
                if d.size > 6291456:
                    return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 6 Mb."})
                if f.is_valid():
                    newfile = None
                    if 'archivo' in request.FILES:
                        newfile = request.FILES['archivo']
                        newfile._name = generar_nombre("informesPer_", newfile._name)
                    informefac = PermisoInstitucional.objects.get(pk=request.POST['id'])
                    informefac.archivoinforme = newfile
                    informefac.save(request)
                    log(u'Subio informes por facultad: %s' % informefac, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                     raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'recalcular':
            try:
                ingreso = IngresoPersonal.objects.get(id=int(request.POST['id']), status=True, estado=1)
                mensaje = ingreso.recalcular()
                if mensaje != '':
                    return JsonResponse({"result": "bad", "mensaje": mensaje})

                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', "mensaje": u"Error al recalcular "})

        elif action == 'recalculamasiva':
            try:
                ingresos = IngresoPersonal.objects.filter(status=True, estado=1)
                for ingreso in ingresos:
                    ingreso.recalcular()
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', "mensaje": u"Error al recalcular "})

        elif action == 'aprorechsolicitud':
            try:
                id = int(encrypt(request.POST['id']))
                estado = int(request.POST['estado'])
                if not SolicitudJustificacionMarcada.objects.filter(status=True, pk=id).exists():
                    JsonResponse({'error': True, 'mensaje': u'No existe solicitud'})
                soli = SolicitudJustificacionMarcada.objects.get(status=True, pk=encrypt(request.POST['id']))
                soli.estado = estado
                soli.save(request)
                log('Edito el estado de la solicitud %s a estado %s' % (soli.__str__(), soli.estado), request, 'edit')
                historial = HistorialSolicitudJustificacionMarcada(
                    solicitud=soli,
                    observacion=request.POST['observacion'],
                    persona=persona,
                    estado=estado,
                    fecha=datetime.now()
                )
                historial.save(request)
                log(u'Agrego historial de solicitud de justificacion de marcadas %s' % (historial.__str__()), request, 'add')
                if int(request.POST['estado'])  == 4:
                    for detalle in soli.detallesolicitudjustificacionmarcada_set.filter(status=True):
                        if detalle.tiposolcitud == 1:
                            marcada = LogMarcada(
                                logdia=detalle.dia,
                                time=detalle.hora,
                                secuencia=detalle.secuencia
                            )
                            marcada.save(request)
                            log('Agrego marcada por justificacion de marcada por omision %s - %s - %s' % (soli.__str__(), detalle.__str__(), marcada.__str__()), request, 'add')
                        elif detalle.tiposolcitud == 2:
                            marcada = detalle.marcada
                            marcada.time = detalle.hora
                            marcada.secuencia = detalle.secuencia
                            marcada.save(request)
                            log('Edito marcada por justificacion de marcada por retraso %s - %s - %s' % (soli.__str__(), detalle.__str__(), marcada.__str__()), request, 'edit')
                    logmarcada = LogMarcada.objects.get(pk=marcada.pk)
                    MarcadasDia.objects.filter(persona=logmarcada.logdia.persona, fecha=logmarcada.logdia.fecha).delete()
                    logmarcada.logdia.cantidadmarcadas = 0
                    logmarcada.logdia.procesado = False
                    logmarcada.logdia.save(request)
                    logmarcada.save(request)
                    cm = logmarcada.logdia.logmarcada_set.filter(status=True).count()
                    logmarcada.logdia.cantidadmarcadas = cm
                    if (cm % 2) == 0:
                        marini = 1
                        for dl in logmarcada.logdia.logmarcada_set.filter(status=True).order_by("time"):
                            if marini == 2:
                                salida = dl.time
                                marini = 1
                                if logmarcada.logdia.persona.marcadasdia_set.filter(fecha=logmarcada.logdia.fecha).exists():
                                    marcadadia = logmarcada.logdia.persona.marcadasdia_set.filter(fecha=logmarcada.logdia.fecha)[0]
                                else:
                                    marcadadia = MarcadasDia(persona=logmarcada.logdia.persona,
                                                             fecha=logmarcada.logdia.fecha,
                                                             logdia=logmarcada.logdia,
                                                             segundos=0)
                                    marcadadia.save(request)
                                if not marcadadia.registromarcada_set.filter(entrada=entrada).exists():
                                    registro = RegistroMarcada(marcada=marcadadia,
                                                               entrada=entrada,
                                                               salida=salida,
                                                               segundos=(salida - entrada).seconds)
                                    registro.save(request)
                                marcadadia.actualizar_marcadas()
                            else:
                                entrada = dl.time
                                marini += 1
                        logmarcada.logdia.procesado = True
                    else:
                        logmarcada.logdia.cantidadmarcadas = 0
                        logmarcada.logdia.procesado = False
                    logmarcada.logdia.save(request)
                    calculando_marcadas(request, logmarcada.time.date(), logmarcada.time.date(), logmarcada.logdia.persona)
                    log(u'editó marcada : %s' % marcada, request, "edit")
                    notificacion(
                        'Solicitud justificación de marcadas',
                        f'Se ha procesado su solicitud justificación de marcadas',
                        soli.solicita,
                        None,
                        'th_hojavida?action=logmarcadas',
                        soli.id,
                        2,
                        'sagest',
                        SolicitudJustificacionMarcada,
                        request
                    )
                return JsonResponse({'error': False})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"error": True, "mensaje": u"Solicitud Incorrecta."})

        elif action == 'subirarchivovacaciones':
            try:
                f = SubirPermisoMasivoForm(request.POST)
                if f.is_valid():
                    nfile = request.FILES['archivo']
                    regimen = f.cleaned_data['regimenlaboral']
                    tipopermiso = f.cleaned_data['tipopermiso']
                    tipopermisodetalle = f.cleaned_data['tipopermisodetalle']
                    motivo = f.cleaned_data['motivo']
                    archivo = io.BytesIO(nfile.read())
                    workbook = load_workbook(filename=archivo, read_only=False)
                    loes = workbook[workbook.sheetnames[0]]
                    ccedula = 1
                    ciniciov = 5
                    cfinv = 6
                    cdia = 7
                    chora = 8
                    cminuto = 9
                    for rowx in range(2, loes.max_row + 1):
                        aprueba = True
                        cedula = loes.cell(row=rowx, column=ccedula).value
                        inicio = loes.cell(row=rowx, column=ciniciov).value
                        fin = loes.cell(row=rowx, column=cfinv).value
                        dia = loes.cell(row=rowx, column=cdia).value
                        hora = loes.cell(row=rowx, column=chora).value
                        minuto = loes.cell(row=rowx, column=cminuto).value
                        distributivo = DistributivoPersona.objects.filter(
                            (Q(persona__cedula__icontains=str(cedula)) | Q(persona__pasaporte__icontains=cedula)),
                            status=True, regimenlaboral=regimen)
                        horainicio = (fin + timedelta(hours=7)).time()
                        horafin = (fin + timedelta(hours=22)).time()
                        concepto = 'Permiso planificado'
                        if distributivo:
                            distributivo = distributivo.first()
                            if not PermisoInstitucionalDetalle.objects.filter(permisoinstitucional__solicita=distributivo.persona, fechainicio=inicio, fechafin=fin,
                                                                          permisoinstitucional__denominacionpuesto=distributivo.denominacionpuesto,
                                                                          permisoinstitucional__unidadorganica=distributivo.unidadorganica,
                                                                          permisoinstitucional__regimenlaboral=regimen,status=True,
                                                                              permisoinstitucional__status=True,permisoinstitucional__estadosolicitud=3).exists():

                                secu = null_to_numeric(
                                    PermisoInstitucional.objects.filter(solicita=distributivo.persona,
                                                                        fechasolicitud__year=datetime.now().year).aggregate(
                                        secu=Max("secuencia"))['secu'])
                                secuencia = secu + 1
                                #
                                permiso = PermisoInstitucional(motivo=motivo, tiposolicitud=1,
                                                               solicita=distributivo.persona,
                                                               fechasolicitud=datetime.now().date(),
                                                               denominacionpuesto=distributivo.denominacionpuesto,
                                                               tipopermiso=tipopermiso,
                                                               tipopermisodetalle=tipopermisodetalle, regimenlaboral=regimen,
                                                               unidadorganica=distributivo.unidadorganica,
                                                               secuencia=secuencia, descuentovacaciones=True,
                                                               estadosolicitud=2)
                                permiso.save()
                                detalle = PermisoInstitucionalDetalle(permisoinstitucional=permiso, fechainicio=inicio,
                                                                      fechafin=fin, horainicio=horainicio, horafin=horafin)
                                detalle.save()

                                kardex = None
                                total = None
                                diassaldo = None
                                horasaldo = None
                                minutosaldo = None
                                i = 1
                                kardex = IngresoPersonal.objects.filter(persona=distributivo.persona, regimenlaboral=regimen,
                                                                        estado=1, status=True)
                                if kardex:
                                    detallekardex = kardex[0].kardexvacacionesdetalle_set.filter(status=True)
                                    if detallekardex:
                                        detallekardex = detallekardex.latest('id')
                                        saldo = timedelta(days=detallekardex.diasal, hours=detallekardex.horasal, minutes=detallekardex.minsal)
                                        permisos = timedelta(days=dia, hours=hora, minutes=minuto)
                                        total=saldo-permisos
                                        if saldo>=permisos:
                                            diassaldo = total.days
                                            horasaldo = total.seconds // 3600
                                            minutosaldo = (total.seconds // 60) - (horasaldo * 60)
                                            detalle = KardexVacacionesDetalle(kardex=kardex[0],
                                                                              permiso=permiso,
                                                                              fecha=datetime.now(),
                                                                              operacion=2,
                                                                              concepto=concepto,
                                                                              diava=dia, horava=hora, minva=minuto,
                                                                              diasal=diassaldo, horasal=horasaldo,
                                                                              minsal=minutosaldo)
                                            detalle.save()
                                        else:
                                            print("no tiene saldo: ", distributivo.persona)
                                            aprueba = False
                                    else:
                                        print("no tiene saldo: ", distributivo.persona)
                                        aprueba = False
                                else:
                                    print("No tiene kardex", rowx, distributivo.persona)
                                    aprueba=False
                                responsabledir = distributivo.unidadorganica.responsable


                                if distributivo.persona == responsabledir or responsabledir is None:
                                    responsabledir = Persona.objects.get(pk=28089)

                                aprobar = PermisoAprobacion(permisoinstitucional=permiso,
                                                            fechaaprobacion=datetime.now().date(),
                                                            observacion="APROBADO",
                                                            aprueba=responsabledir,
                                                            estadosolicitud=1)
                                aprobar.save()
                                if aprueba:
                                    aprobar = PermisoAprobacion(permisoinstitucional=permiso,
                                                                fechaaprobacion=datetime.now().date(),
                                                                observacion="APROBADO",
                                                                aprueba_id=122274,
                                                                estadosolicitud=1)
                                    aprobar.save()

                                    permiso.estadosolicitud=3
                                    permiso.save()

                    return JsonResponse({"result": False}, safe=False)
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, 'mensaje': str(ex)}, safe=False)


        return JsonResponse({"result": "bad", "mensaje": u"Solicitud Incorrecta."})

    else:
        if 'action' in request.GET:
            data['action'] = action = request.GET['action']

            if action == 'verdetalle':
                try:
                    data = {}
                    detalle = PermisoInstitucional.objects.get(pk=int(request.GET['id']))
                    data['permiso'] = detalle
                    data['detallepermiso'] = detalle.permisoinstitucionaldetalle_set.all()
                    data['aprobadores'] = detalle.permisoaprobacion_set.all()
                    template = get_template("th_permiso_institucional/detalle.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            # if action == 'verdetalle':
            #     try:
            #         data = {}
            #         detalle = PermisoInstitucional.objects.get(pk=int(request.GET['id']))
            #         data['permiso'] = detalle
            #         data['detallepermiso'] = detalle.permisoinstitucionaldetalle_set.all()
            #         data['aprobadores'] = detalle.permisoaprobacion_set.all()
            #         template = get_template("th_permiso_institucional/detalle.html")
            #         json_content = template.render(data)
            #         return JsonResponse({"result": "ok", 'data': json_content})
            #     except Exception as ex:
            #         return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            elif action == 'xlsaprobarpermiso':
                try:
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=listado_aprobarpermiso.xls'
                    wb = xlwt.Workbook()
                    ws = wb.add_sheet('Sheetname')
                    estilo = xlwt.easyxf('font: height 350, name Arial, colour_index black, bold on, italic on; align: wrap on, vert centre, horiz center;')
                    ws.col(0).width = 1000
                    ws.col(1).width = 6000
                    ws.col(2).width = 3000
                    ws.col(3).width = 3000
                    ws.col(4).width = 3000
                    ws.col(5).width = 8000
                    ws.col(6).width = 8000
                    ws.col(7).width = 3000
                    ws.col(8).width = 3000
                    ws.col(9).width = 15000
                    ws.col(10).width = 2000
                    ws.col(11).width = 15000
                    ws.col(12).width = 2000
                    ws.col(20).width = 5000
                    ws.col(21).width = 5000

                    ws.write(0, 0, 'N.')
                    ws.write(0, 1, 'CODIGO')
                    ws.write(0, 2, 'FECHA')
                    ws.write(0, 3, 'FECHA INICIO')
                    ws.write(0, 4, 'FECHA FIN')
                    ws.write(0, 5, 'HORA INICIO')
                    ws.write(0, 6, 'HORA FIN')
                    ws.write(0, 7, 'DIAS')
                    ws.write(0, 8, 'HORAS')
                    ws.write(0, 9, 'ESTADO')
                    ws.write(0, 10, 'CEDULA')
                    ws.write(0, 11, 'SOLICITANTE')
                    ws.write(0, 12, 'DEPARTAMENTO')
                    ws.write(0, 13, 'TIPO SOLICIDUTD')
                    ws.write(0, 14, 'TIPO PERMISO')
                    ws.write(0, 15, 'DETALLE PERMISO')
                    ws.write(0, 16, 'DESCUENTO VACACIONES')
                    ws.write(0, 17, 'MOTIVO')
                    ws.write(0, 18, 'SOPORTE')
                    ws.write(0, 19, 'USUARIO')
                    ws.write(0, 20, 'CASA SALUD')
                    ws.write(0, 21, 'REGIMEN LABORAL')

                    a = 0
                    date_format = xlwt.XFStyle()
                    fechainicio = request.GET['fechainicio']
                    fechafinal = request.GET['fechafinal']
                    date_format.num_format_str = 'yyyy/mm/dd'
                    plantillas = PermisoInstitucional.objects.filter(fechasolicitud__gte=fechainicio, fechasolicitud__lte=fechafinal).order_by('estadosolicitud', '-fechasolicitud')
                    archivos = ''
                    for per in plantillas:
                        nompersona = ''
                        if PermisoAprobacion.objects.filter(permisoinstitucional=per).exists():
                            aprobadores = PermisoAprobacion.objects.filter(permisoinstitucional=per)
                            if aprobadores.values('id').count() > 1:
                                apro = PermisoAprobacion.objects.filter(permisoinstitucional=per).order_by('-id')[0]
                                nompersona = u"%s" % apro.aprueba
                        listafini = []
                        listaffin = []
                        listahini = []
                        listahfin = []
                        diastotales = ''
                        horastotales = ''
                        permisosdetalles = PermisoInstitucionalDetalle.objects.filter(permisoinstitucional=per)
                        for perdetalles in permisosdetalles:
                            if permisosdetalles.values('id').count() > 1:
                                listafini.append(('%s-' % perdetalles.fechainicio))
                                listaffin.append(('%s-' % perdetalles.fechafin))
                                listahini.append(('%s-' % perdetalles.horainicio))
                                listahfin.append(('%s-' % perdetalles.horafin))
                            else:
                                cursor = connection.cursor()
                                cursor.execute("select CAST((fechafin-fechainicio)AS text) as dias,CAST((horafin-horainicio)AS text) as horas from sagest_PermisoInstitucionalDetalle where id="+str(perdetalles.id))
                                results = cursor.fetchall()
                                diastotales = 0;
                                for r in results:
                                    if r[0] == '0':
                                        diastotales = r[0]
                                    else:
                                        diastotales = int(r[0])+1
                                    horastotales = r[1]
                                fini = perdetalles.fechainicio
                                ffin = perdetalles.fechafin

                                listahini.append(str(perdetalles.horainicio))
                                listahfin.append(str(perdetalles.horafin))
                        a += 1
                        ws.write(a, 0, a)
                        ws.write(a, 1, per.codificacion())
                        ws.write(a, 2, per.fechasolicitud, date_format)
                        if permisosdetalles.values('id').count() == 1:
                            ws.write(a, 3, fini, date_format)
                            ws.write(a, 4, ffin, date_format)
                            ws.write(a, 5, listahini)
                            ws.write(a, 6, listahfin)
                        else:
                            ws.write(a, 3, '%s' % listafini)
                            ws.write(a, 4, '%s' % listaffin)
                            ws.write(a, 5, '%s' % listahini)
                            ws.write(a, 6, '%s' % listahfin)
                        ws.write(a, 7, diastotales)
                        ws.write(a, 8, horastotales)
                        ws.write(a, 9, per.get_estadosolicitud_display())
                        ws.write(a, 10, per.solicita.cedula)
                        ws.write(a, 11, u"%s" % per.solicita.nombre_completo())
                        ws.write(a, 12, u"%s" % per.unidadorganica.nombre)
                        ws.write(a, 13, per.get_tiposolicitud_display().upper())
                        if per.tipopermiso:
                            ws.write(a, 14, u"%s" % per.tipopermiso.descripcion)
                        else:
                            ws.write(a, 14, '')
                        if per.tipopermisodetalle:
                            ws.write(a, 15, u"%s" % per.tipopermisodetalle.descripcion)
                        else:
                            ws.write(a, 15, '')
                        if per.descuentovacaciones:
                            ws.write(a, 16, 'SI')
                        else:
                            ws.write(a, 16, 'NO')
                        ws.write(a, 17, u"%s" % per.motivo)
                        if per.archivo:
                            archivos = 'SI'
                        else:
                            archivos = 'NO'
                        ws.write(a, 18, archivos)
                        ws.write(a, 19, nompersona)
                        ws.write(a, 20, str(per.casasalud.descripcion) if per.casasalud else "")
                        ws.write(a, 21, str(per.regimenlaboral) if per.regimenlaboral else "")
                    wb.save(response)
                    return response
                except Exception as ex:
                    pass

            if action == 'detalle':
                try:
                    data = {}
                    dias=0
                    horas=0
                    minutos=0
                    validacion = int(request.GET['validacion'])
                    data['permisos_salud_ocupa'] = permisos_salud_ocupa = True if request.user.has_perm('sagest.ver_permisos_salud_ocupa') and not request.user.is_superuser else False
                    detalle = PermisoInstitucional.objects.get(pk=int(request.GET['id']))
                    data['permiso'] = detalle
                    data['detallepermiso'] =detallepermiso= detalle.permisoinstitucionaldetalle_set.all()
                    data['aprobador'] = persona
                    data['th'] = 0 if permisos_salud_ocupa else 1
                    data['descuentovacaciones'] = detalle.descuentovacaciones
                    data['fecha'] = datetime.now().date()
                    data['aprobadores'] = detalle.permisoaprobacion_set.all()
                    i=1
                    for d in detallepermiso:
                        if d.horainicio:
                            if i == 1:
                                inicio=timedelta( hours=d.horainicio.hour, minutes=d.horainicio.minute)
                                fin=timedelta( hours=d.horafin.hour, minutes=d.horafin.minute)
                                total=fin-inicio
                                fechaprimerini = d.fechainicio
                                fechaprimerfin = d.fechafin
                                # horas = d.horafin.hour - d.horainicio.hour
                                # minutos = d.horafin.minute - d.horainicio.minute
                                horas = total.seconds // 3600
                                minutos = (total.seconds // 60) - (horas * 60)
                                if d.fechafin != d.fechainicio:
                                    dias = d.fechafin - d.fechainicio
                                    dias = dias.days
                                if minutos >=60:
                                    minutos=0
                                    horas=horas+1
                                if horas>=8:
                                    dias=dias+1
                                    horas=0
                                i = i + 1
                            else:
                                # horas = horas + (d.horafin.hour - d.horainicio.hour)
                                # minutos = minutos + (d.horafin.minute - d.horainicio.minute)
                                inicio = timedelta(hours=d.horainicio.hour, minutes=d.horainicio.minute)
                                fin = timedelta(hours=d.horafin.hour, minutes=d.horafin.minute)
                                total1 = fin - inicio
                                horas =horas+ (total1.seconds // 3600)
                                minutos =minutos+((total1.seconds // 60) - ((total1.seconds // 3600) * 60))
                                if fechaprimerini != d.fechainicio and fechaprimerfin != d.fechafin:
                                    resta=(d.fechafin - d.fechainicio)
                                    dias = dias + resta.days
                                    # dias=dias.days
                                if minutos >=60:
                                    minutos=0
                                    horas=horas+1
                                if horas >= 8:
                                    dias = dias + 1
                                    horas = 0

                    data['dias']=dias
                    data['horas'] = horas
                    data['minutos'] = minutos
                    if validacion == 0:
                        template = get_template("th_permiso_institucional/detalle_aprobar.html")
                    else:
                        template = get_template("th_permiso_institucional/detalle_rechazar.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            elif action == 'add':
                try:
                    data['title'] = u'Agregar Permiso Institucional'
                    form = PermisoInstitucionalAdicionarForm(initial={'fechasolicitud': datetime.now().date()})
                    form.adicionar()
                    # form.deshabilitar_permisofamilia()
                    data['form2'] = PermisoInstitucionalDetalleForm()
                    data['fecha'] = str(datetime.now().date())[8:10] + '-' + str(datetime.now().date())[5:7] + '-' + str(datetime.now().date())[:4]
                    data['hora'] = str(datetime.now())[11:16]
                    data['form'] = form
                    return render(request, "th_permiso_institucional/addadmin.html", data)
                except Exception as ex:
                    pass

            elif action == 'kardexvacaciones':
                try:
                    data['title'] = u'Kardex Vacaciones'
                    search = None
                    ids = None
                    nombramientoid = None
                    regimenid = None
                    if 's' in request.GET:
                        search = request.GET['s']
                        if not search.isdigit():
                            ss = search.split(' ')
                            if len(ss) == 1:
                                kardex = IngresoPersonal.objects.filter(Q(persona__apellido1__icontains=search)|Q(persona__apellido2__icontains=search), status=True).order_by("-fechaingreso")
                            else:
                                kardex = IngresoPersonal.objects.filter(Q(persona__apellido1__icontains=ss[0]) & Q(persona__apellido2__icontains=ss[1]), status=True).order_by("-fechaingreso")
                        else:
                            kardex = IngresoPersonal.objects.filter(Q(persona__cedula__icontains=search) , status=True).order_by("-fechaingreso")
                    elif 'id' in request.GET:
                        ids = request.GET['id']
                        kardex = IngresoPersonal.objects.filter(status=True,id=ids).order_by("-fechaingreso")
                    elif 'regimenid' in request.GET and 'nombramientoid' in request.GET:
                        regimenid = int(request.GET['regimenid'])
                        nombramientoid = int(request.GET['nombramientoid'])
                        if nombramientoid == 1:
                            kardex = IngresoPersonal.objects.filter(status=True,regimenlaboral_id=regimenid ,nombramiento=True).order_by("-fechaingreso")
                        else:
                            kardex = IngresoPersonal.objects.filter(status=True,regimenlaboral_id=regimenid, nombramiento=False).order_by("-fechaingreso")
                    elif 'regimenid' in request.GET:
                        regimenid = int(request.GET['regimenid'])
                        kardex = IngresoPersonal.objects.filter(status=True,regimenlaboral_id=regimenid).order_by("-fechaingreso")
                    elif 'nombramientoid' in request.GET:
                        nombramientoid = int(request.GET['nombramientoid'])
                        if nombramientoid == 1:
                            kardex = IngresoPersonal.objects.filter(status=True,nombramiento=True).order_by("-fechaingreso")
                        else:
                            kardex = IngresoPersonal.objects.filter(status=True, nombramiento=False).order_by("-fechaingreso")
                    else:
                        kardex = IngresoPersonal.objects.filter(status=True).order_by("persona")
                    paging = MiPaginador(kardex, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['page'] = page
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['kardex'] = page.object_list
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['regimenid'] = regimenid if regimenid else ""
                    data['nombramientoid'] = nombramientoid if nombramientoid else ""
                    data['regimen']= RegimenLaboral.objects.filter(status=True)
                    return render(request, "th_permiso_institucional/kardexvacaciones.html", data)
                except Exception as ex:
                    pass

            elif action == 'descargarkardex':
                try:
                    __author__ = 'Unemi'
                    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
                    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                    title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    style1 = easyxf(num_format_str='D-MMM-YY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('exp_xls_post_part')
                    ws.write_merge(0, 0, 0, 7, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=empleados_' + random.randint(1, 10000).__str__() + '.xls'
                    columns = [
                        (u"N", 1000),
                        (u"PERSONA", 11000),
                        (u"REGIMEN", 4000),
                        (u"FECHA INGRESO", 4000),
                        (u"FECHA SALIDA", 4000),
                        (u"ESTADO", 4000),
                        (u"SALDO", 4000),
                    ]
                    row_num = 3
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    cursor = connection.cursor()
                    listadokardex = IngresoPersonal.objects.filter(status=True).order_by("persona")
                    row_num = 4
                    a = 0
                    for kar in listadokardex:
                        i = 0
                        campo1 = kar.persona.apellido1 + ' ' + kar.persona.apellido2 + ' ' + kar.persona.nombres
                        campo2 = kar.regimenlaboral.descripcion
                        campo3 = kar.fechaingreso
                        campo4 = kar.get_estado_display()
                        campo5 = kar.saldo()
                        a += 1
                        ws.write(row_num, 0, a, font_style2)
                        ws.write(row_num, 1, campo1, font_style2)
                        ws.write(row_num, 2, campo2, font_style2)
                        ws.write(row_num, 3, campo3, style1)
                        if kar.fechasalida:
                            ws.write(row_num, 4, kar.fechasalida, style1)
                        else:
                            ws.write(row_num, 4, 'Actualidad', font_style2)
                        ws.write(row_num, 5, campo4, font_style2)
                        ws.write(row_num, 6, campo5, font_style2)
                        row_num += 1
                    wb.save(response)
                    connection.close()
                    return response
                except Exception as ex:
                    pass

            elif action == 'addkardexindividual':
                try:
                    data['title'] = u'Agregar '
                    form= KardexVacacionesIndividualForm()
                    form.adicionar()
                    data['form'] = form
                    return render(request, "th_permiso_institucional/addkardexindividual.html", data)
                except Exception as ex:
                    pass

            elif action == 'addkardex':
                try:
                    data['title'] = u'Agregar Kardex'
                    data['form']=form= KardexVacacionesForm()
                    return render(request, "th_permiso_institucional/addkardex.html", data)
                except Exception as ex:
                    pass

            elif action == 'editkardex':
                try:
                    data['title'] = u'Editar'
                    data['informe'] = kardex = IngresoPersonal.objects.get(pk=int(request.GET['id']))
                    form= KardexVacacionesIndividualForm(initial={'fechaingreso': kardex.fechaingreso,
                                                                  'fechasalida': kardex.fechasalida,
                                                                  'estado': kardex.estado,
                                                                  'regimenlaboral':kardex.regimenlaboral,
                                                                  'nombramiento':kardex.nombramiento,
                                                                  'contratoindefinido': kardex.contratoindefinido})
                    form.editar(kardex)
                    data['form'] = form
                    return render(request, "th_permiso_institucional/editkardex.html", data)
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos. %s" % ex})
                    pass

            elif action == 'delekardex':
                try:
                    data['title'] = u'Eliminar Kardex'
                    data['kardex'] = IngresoPersonal.objects.get(pk=int(request.GET['id']))
                    return render(request, "th_permiso_institucional/delekardex.html", data)
                except Exception as ex:
                    pass

            elif action == 'detallekardex':
                try:
                    search=None
                    ids=None
                    data['id']= id =int(request.GET['id'])
                    data['kardex']=kardex=IngresoPersonal.objects.get(pk=int(request.GET['id']))
                    detalle=kardex.kardexvacacionesdetalle_set.filter(status=True).order_by('-fecha')
                    data['title'] = 'Detalle Kardex de Vacaciones'

                    paging = MiPaginador(detalle, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['page'] = page
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['detalle'] = page.object_list
                    data['url_vars'] = "&action=%s&id=%s" % (action,id)
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    return render(request, 'th_permiso_institucional/detallekardex.html', data)
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            elif action == 'delkardex':
                try:
                    data['title'] = u'Eliminar Kardex'
                    data['kardex'] = KardexVacacionesDetalle.objects.get(pk=int(request.GET['id']))
                    return render(request, "th_permiso_institucional/delkardex.html", data)
                except Exception as ex:
                    pass


            # elif action == 'addkardexindividualdetalle':
            #     try:
            #         data['id']=int(request.GET['id'])
            #         data['idp']=int(request.GET['idp'])
            #         data['regimen']=int(request.GET['regimen'])
            #         persona=Persona.objects.get(pk=int(request.GET['idp']))
            #         data['title'] = u'AGREGAR KARDEX '+str(persona.nombre_completo_inverso())
            #         form=KardexVacacionesDetalleForm(initial={'operacion': 2})
            #         form.adicionardetalle(persona)
            #         data['form'] =form
            #         return render(request, "th_permiso_institucional/addkardexindividualdeta.html", data)
            #     except Exception as ex:
            #         pass

            if action == 'addkardexindividualdetalle':
                try:
                    data['filtro']=IngresoPersonal.objects.get(id=int(request.GET['id']))
                    form = KardexVacacionesDetalleForm(initial={'operacion': 1})
                    form.adicionardetalle(persona)
                    data['form'] = form
                    template = get_template("th_permiso_institucional/addkardexindividualdeta.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass


            elif action == 'editfecha':
                try:
                    data['title'] = u'Solicitar Permiso Institucional'
                    data['permiso'] = permiso = PermisoInstitucional.objects.get(pk=int(request.GET['id']))
                    distributivo = DistributivoPersona.objects.filter(persona=permiso.solicita,denominacionpuesto=permiso.denominacionpuesto, status=True,estadopuesto__id=PUESTO_ACTIVO_ID).distinct()
                    denominacionpuesto = None
                    if distributivo.exists():
                        denominacionpuesto = distributivo[0].denominacionpuesto
                    form = PermisoInstitucionalFechaForm(initial={'fechasolicitud': permiso.fechasolicitud,
                                                                  'tiposolicitud': permiso.tiposolicitud,
                                                                  'tipopermiso': permiso.tipopermiso,
                                                                  'tipopermisodetalle': permiso.tipopermisodetalle,
                                                                  'denominacionpuesto': denominacionpuesto,
                                                                  'motivo': permiso.motivo,
                                                                  'casasalud':permiso.casasalud})
                    form.editar_aprobar(permiso.solicita)
                    data['form2'] = PermisoInstitucionalDetalleForm()
                    data['fecha'] = str(datetime.today())[8:10] + '-' + str(datetime.today())[5:7] + '-' + str(datetime.today())[:4]
                    data['hora'] = str(datetime.today())[11:16]
                    data['form'] = form
                    data['puedesubirarchivo'] = 1 if permiso.estadosolicitud in [1, 2, 3] and not permiso.archivo else 0
                    data['detalles'] = permiso.permisoinstitucionaldetalle_set.all()
                    return render(request, "th_permiso_institucional/editfecha.html", data)
                except Exception as ex:
                    pass

            elif action == 'subirdocumentosoporte':
                try:
                    data['id'] = int(request.GET['id'])
                    data['form'] = InformesPermisoForm
                    template = get_template('ajaxformmodal.html')
                    return JsonResponse({'result': True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({'result': False, 'mensaje': str(ex)}, safe=False)

            elif action == 'addsubirinforme':
                try:
                    data['title'] = u'Informe'
                    data['form'] = InformesPermisoForm
                    data['permisoinstitucional'] = PermisoInstitucional.objects.get(pk=request.GET['id'])
                    template = get_template("th_permiso_institucional/add_informepermiso.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content})
                except Exception as ex:
                    pass

            elif action == 'recalculamasiva':
                try:
                    data['title'] = u'Recalcular Kardex'
                    return render(request, "th_permiso_institucional/recalculamasiva.html", data)
                except Exception as ex:
                    pass
            if action =='justifiacionmarcadas':
                try:
                    data['title'] = u'Solicitud de justificación de marcadas'
                    estado = [1]
                    url_vars = ''
                    if request.user.has_perm('sagest.puede_aprobar_justificacion_marcada_director'):
                        data['es_director_th'] = True
                        estado = [2, 3, 4, 5]
                    elif request.user.has_perm('sagest.puede_cerrar_justificacion_marcada_analista'):
                        data['es_analista_th'] = True
                        estado = [3, 4, 5]
                    solicitud =SolicitudJustificacionMarcada.objects.filter(status=True, estado__in=estado)
                    if 's' in request.GET:
                        data['s'] = search = request.GET['s']
                        ss = search.split(' ')
                        if len(ss) > 1:
                            solicitud = solicitud.filter(Q(solicita__apellido1__icontains=ss[0]) & Q(solicita__apellido2__icontains=ss[1]))
                        else:
                            solicitud = solicitud.filter(Q(id__icontains=search) | Q(solicita__apellido1__icontains=search) | Q(solicita__cedula__icontains=search)).distinct()
                        url_vars += f"&s={search}"
                    if 'estado' in request.GET:
                        data['estado'] = estado = int(request.GET['estado'])
                        solicitud = solicitud.filter(estado=estado)
                        url_vars += f"&estado={estado}"
                    paging = MiPaginador(solicitud, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['solicitudes'] = page.object_list
                    data['url_vars'] = url_vars
                    return render(request,"th_permiso_institucional/solicitudmarcadas.html",data)
                except Exception as ex:
                    pass

            if action == 'subirarchivovacaciones':
                try:
                    data['form'] = SubirPermisoMasivoForm()
                    data['action'] = 'subirarchivovacaciones'
                    template = get_template("th_permiso_institucional/subir_vacaciones.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'mensaje': str(ex)}, safe=False)

            return HttpResponseRedirect(request.path)

        else:
            try:
                data['title'] = u'Permisos institucionales'
                search = None
                excluir = []
                permisos_salud_ocupa = False
                estados, search, url_vars = request.GET.get('estados', ''),request.GET.get('s', ''), ''

                if not request.user.is_superuser:
                    # if request.user.has_perm('sagest.ver_permisos_auth'):
                        # excluir = [x.id for x in TipoPermiso.objects.filter(status=True, quienaprueba=2)]

                    if request.user.has_perm('sagest.ver_permisos_salud_ocupa'):
                        excluir = [x.id for x in TipoPermiso.objects.filter(status=True, quienaprueba=1)]
                        permisos_salud_ocupa = True

                ids = None
                if 's' in request.GET:
                    data['s'] = search = request.GET['s'].strip()
                    url_vars += f"&s={search}"
                    ss = search.split(' ')
                    if len(ss) == 1:
                        plantillas = PermisoInstitucional.objects.filter(Q(status=True), Q(solicita__nombres__icontains=search) |
                                                                         Q(solicita__apellido1__icontains=search) |
                                                                         Q(solicita__apellido2__icontains=search) |
                                                                         Q(solicita__cedula__icontains=search) |
                                                                         Q(solicita__pasaporte__icontains=search)).exclude(tipopermiso__in=excluir).distinct().order_by('-fechasolicitud')
                    else:
                        plantillas = PermisoInstitucional.objects.filter(Q(status=True), Q(solicita__apellido1__icontains=ss[0]) & Q(solicita__apellido2__icontains=ss[1])).exclude(tipopermiso__in=excluir).distinct().order_by('-fechasolicitud')
                else:
                    if permisos_salud_ocupa:
                        plantillas = PermisoInstitucional.objects.filter(status=True).exclude(tipopermiso__in=excluir).order_by('estadosolicitud', '-fechasolicitud')
                    else:
                        if persona.es_directordepartamental() or persona.usuario.is_superuser:
                            plantillas = PermisoInstitucional.objects.filter(status=True).order_by('estadosolicitud', '-fechasolicitud')
                        else:
                            excluir2 = [x.id for x in TipoPermiso.objects.filter(status=True, quienaprueba=2)]

                            plantillas = PermisoInstitucional.objects.filter(status=True).exclude(tipopermiso__in=excluir2).order_by('estadosolicitud',
                                                                                           '-fechasolicitud')
                if estados:
                    data['estados'] = int(estados)
                    plantillas= plantillas.filter(estadosolicitud=estados)
                    url_vars += f"&estados={estados}"

                paging = MiPaginador(plantillas, 20)
                p = 1
                try:
                    paginasesion = 1
                    if 'paginador' in request.session:
                        paginasesion = int(request.session['paginador'])
                    if 'page' in request.GET:
                        p = int(request.GET['page'])
                    else:
                        p = paginasesion
                    try:
                        page = paging.page(p)
                    except:
                        p = 1
                    page = paging.page(p)
                except:
                    page = paging.page(p)
                request.session['paginador'] = p
                data['paging'] = paging
                data['rangospaging'] = paging.rangos_paginado(p)
                data['page'] = page
                data['estadosp'] = ESTADO_PERMISOS
                data['search'] = search if search else ""
                data['ids'] = ids if ids else ""
                data['permisos'] = page.object_list
                data['url_vars'] = url_vars
                data['email_domain'] = EMAIL_DOMAIN
                data['permisos_salud_ocupa'] = permisos_salud_ocupa
                return render(request, 'th_permiso_institucional/aprobar_th.html', data)
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})
def calculo_fecha_dias_plazo(diasplazo, dias_antes_despues, fechaempezar=None):
    if fechaempezar:
        fecha = fechaempezar
    else:
        fecha = datetime.now().date()
    aumento = 1
    while aumento <= diasplazo:
        dias = None
        if dias_antes_despues == 1:
            dias = timedelta(days=1)
        elif dias_antes_despues == 2:
            dias = timedelta(days=-1)
        fecha = fecha + dias
        aumento += 1
    return fecha.strftime('%d-%m-%Y')

def extraer_regimen_laboral_antes_guardar(distributivo_id):
    distributivo = DistributivoPersona.objects.filter(pk=distributivo_id)
    if distributivo.exists():
        return distributivo[0].regimenlaboral
    else:
        return None


def valida_este_en_fecha(tipopermisodetalle, fechadesde, fechasolicitud=None):
    mensaje = ''
    puedeadicionar = False
    fecha = datetime.strptime(calculo_fecha_dias_plazo(tipopermisodetalle.diasplazo, tipopermisodetalle.aplicar, fechasolicitud if fechasolicitud else None), '%d-%m-%Y').date()
    if tipopermisodetalle.aplicar == 1:
        if fechadesde >= fecha:
            puedeadicionar = True
        else:
            puedeadicionar = False
            mensaje = "DEBERÁ SOLICITAR  " + ('HASTA ' if tipopermisodetalle.aplicar == 2 else '') + tipopermisodetalle.diasplazo.__str__() +' DÍAS '+ tipopermisodetalle.get_aplicar_display().__str__()
    elif tipopermisodetalle.aplicar == 2:
        if fechadesde >= fecha:
            puedeadicionar = True
        else:
            puedeadicionar = False
            mensaje = "DEBERÁ SOLICITAR  " + ('HASTA ' if tipopermisodetalle.aplicar == 2 else '') + tipopermisodetalle.diasplazo.__str__() + ' DÍAS ' + tipopermisodetalle.get_aplicar_display().__str__()
    return [puedeadicionar, mensaje, fecha]


def valida_este_en_dias(tipopermisodetalle, fechainicio, fechafin):
    fechaini = fechainicio
    contador = 0
    mensaje = ''
    puedeadicionar = False
    while fechaini <= fechafin:
        fechaini = fechaini + timedelta(days=1)
        contador += 1
    if contador <= tipopermisodetalle.dias:
        puedeadicionar = True
    else:
        mensaje = "EL PERMISO NO PUEDE EXCEDER MÁS DE "+tipopermisodetalle.dias.__str__()+" DÍAS"
    return [puedeadicionar, mensaje]

